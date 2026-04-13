#\!/usr/bin/env python3
"""
Genera FASE_7_CARGAR_FORMULAS_LIMPIAS.sql
Mapea ingredientes de Formulas Maestras a códigos MPMP normalizados
"""

import pandas as pd
from pathlib import Path
import openpyxl
import re

print("=" * 80)
print("GENERAR FASE 7: CARGAR FORMULAS CON CÓDIGOS MPMP NORMALIZADOS")
print("=" * 80)

# 1. RECONSTRUIR MAPEO MATERIAL -> MPMP CODE
print("\n1️⃣  Reconstruyendo mapeo de códigos MPMP...")

mapeo = {}  # material_name -> mpmp_code
mpmp_counter = 1

# Leer inventario
inv_file = Path("inventario-espagiria/INVENTARIO REAL DE MATERIAS PRIMAS.xlsx")
try:
    df_inv = pd.read_excel(inv_file)
    mat_col = None
    for col in df_inv.columns:
        if 'materia' in col.lower() or 'material' in col.lower() or 'nombre' in col.lower():
            mat_col = col
            break
    if not mat_col:
        mat_col = df_inv.columns[0]
    
    inv_materiales = set()
    for mat in df_inv[mat_col].dropna():
        mat_clean = str(mat).strip().upper()
        if mat_clean:
            inv_materiales.add(mat_clean)
except Exception as e:
    print(f"   ✗ Error leyendo inventario: {e}")
    inv_materiales = set()

# Leer formulas
formula_materiales = set()
base = Path("Formulas Maestras")
for prod_folder in sorted(base.iterdir()):
    if not prod_folder.is_dir():
        continue
    for excel_file in prod_folder.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            if 'OP' in wb.sheetnames:
                ws = wb['OP']
                for row in range(15, 100):
                    val = ws[f'B{row}'].value
                    if val:
                        mat = str(val).strip().upper()
                        formula_materiales.add(mat)
                    elif val is None:
                        break
            wb.close()
        except:
            pass

# Consolidar y asignar códigos
materiales_consolidados = sorted(inv_materiales | formula_materiales)
for mat in materiales_consolidados:
    mapeo[mat] = f"MPMP{mpmp_counter:05d}"
    mpmp_counter += 1

print(f"   ✓ Mapeo creado: {len(mapeo)} materiales")

# 2. LEER FORMULAS MAESTRAS Y EXTRAER INGREDIENTES
print("\n2️⃣  Extrayendo fórmulas maestras con ingredientes...")

formulas = {}  # {nombre_producto: [{ingrediente: nombre, cantidad: qty, unidad: unit, ...}]}

for prod_folder in sorted(base.iterdir()):
    if not prod_folder.is_dir():
        continue
    
    prod_name = prod_folder.name
    
    for excel_file in prod_folder.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            if 'OP' in wb.sheetnames:
                ws = wb['OP']
                
                ingredientes = []
                for row in range(15, 100):
                    nombre_cell = ws[f'B{row}'].value
                    cantidad_cell = ws[f'C{row}'].value
                    
                    if nombre_cell and nombre_cell.strip():
                        nombre = str(nombre_cell).strip().upper()
                        cantidad = cantidad_cell if cantidad_cell else 0
                        
                        # Obtener código MPMP
                        mpmp_code = mapeo.get(nombre, None)
                        
                        if mpmp_code:
                            ingredientes.append({
                                'nombre': nombre,
                                'cantidad': cantidad,
                                'mpmp_code': mpmp_code
                            })
                    elif nombre_cell is None:
                        break
                
                if ingredientes:
                    formulas[prod_name] = ingredientes
                    print(f"   ✓ {prod_name}: {len(ingredientes)} ingredientes")
            
            wb.close()
        except Exception as e:
            print(f"   ⚠️  Error en {excel_file.name}: {e}")

print(f"\n   Total formulas maestras: {len(formulas)}")

# 3. GENERAR FASE_7_CARGAR_FORMULAS_LIMPIAS.sql
print("\n3️⃣  Generando FASE_7_CARGAR_FORMULAS_LIMPIAS.sql...")

sql_inserts = []
for producto, ingredientes in formulas.items():
    for ing in ingredientes:
        sql_line = f"  ('{ing['mpmp_code']}', '{producto}', {ing['cantidad']})"
        sql_inserts.append(sql_line)

sql_content = f"""-- ============================================================================
-- FASE 7: CARGAR FORMULAS LIMPIAS CON CÓDIGOS MPMP NORMALIZADOS
-- ============================================================================
-- Mapeo de ingredientes a códigos MPMP
-- Fuente: Formulas Maestras + Materiales Consolidados
-- Total: {len(sql_inserts)} relaciones formula-ingrediente
-- ============================================================================

INSERT INTO formulas_productos (codigo_material, nombre_producto, cantidad)
VALUES
""" + ",\n".join(sql_inserts) + ";"

with open("FASE_7_CARGAR_FORMULAS_LIMPIAS.sql", "w", encoding="utf-8") as f:
    f.write(sql_content)

print(f"\n✅ ARCHIVO GENERADO:")
print(f"   • FASE_7_CARGAR_FORMULAS_LIMPIAS.sql")
print(f"\n📊 ESTADÍSTICAS:")
print(f"   • Materiales mapeados: {len(mapeo)}")
print(f"   • Formulas cargadas: {len(formulas)}")
print(f"   • Relaciones ingrediente-formula: {len(sql_inserts)}")
print("=" * 80)
