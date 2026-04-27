#!/usr/bin/env python3
"""
Cruza inventario real con formulas maestras
Genera lista maestra consolidada con códigos MPMP
"""

import pandas as pd
from pathlib import Path
import openpyxl

print("=" * 80)
print("CRUZAR INVENTARIO REAL vs FORMULAS MAESTRAS")
print("=" * 80)

# 1. LEER INVENTARIO REAL
print("\n1️⃣  Leyendo INVENTARIO REAL DE MATERIAS PRIMAS.xlsx...")
inv_file = Path("inventario-espagiria/INVENTARIO REAL DE MATERIAS PRIMAS.xlsx")

try:
    df_inv = pd.read_excel(inv_file)
    print(f"   ✓ Cargado: {len(df_inv)} filas")
    print(f"   Columnas: {list(df_inv.columns)}")
    print(f"\n   Primeras filas:")
    print(df_inv.head())
except Exception as e:
    print(f"   ✗ Error: {e}")
    exit(1)

# 2. EXTRAER MATERIALES DEL INVENTARIO
print("\n2️⃣  Extrayendo materiales del inventario...")

# Identificar columna con nombre de material (buscar por nombre común)
mat_col = None
for col in df_inv.columns:
    if 'materia' in col.lower() or 'material' in col.lower() or 'nombre' in col.lower():
        mat_col = col
        break

if not mat_col:
    mat_col = df_inv.columns[0]  # Usar primera columna por defecto
    print(f"   ⚠️  Usando columna por defecto: {mat_col}")
else:
    print(f"   ✓ Columna de materiales: {mat_col}")

inv_materiales = set()
for mat in df_inv[mat_col].dropna():
    mat_clean = str(mat).strip().upper()
    if mat_clean:
        inv_materiales.add(mat_clean)

print(f"   ✓ Materiales en inventario: {len(inv_materiales)}")
print(f"\n   Muestra (primeros 15):")
for i, mat in enumerate(sorted(inv_materiales)[:15], 1):
    print(f"      {i:2d}. {mat}")

# 3. EXTRAER MATERIALES DE FORMULAS MAESTRAS
print("\n3️⃣  Extrayendo materiales de Formulas Maestras...")

formula_materiales = set()
base = Path("Formulas Maestras")

for prod_folder in sorted(base.iterdir()):
    if not prod_folder.is_dir():
        continue

    for excel_file in prod_folder.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            if 'OP' in wb.sheetnames:
                ws = wb['OP']
                for row in range(15, 100):
                    val = ws[f'B{row}'].value
                    if val:
                        mat = str(val).strip().upper()
                        formula_materiales.add(mat)
                    elif val is None:
                        break
            wb.close()
        except:
            pass

print(f"   ✓ Materiales en formulas: {len(formula_materiales)}")

# 4. CRUCE
print("\n4️⃣  Analizando diferencias...")

solo_en_inventario = inv_materiales - formula_materiales
solo_en_formulas = formula_materiales - inv_materiales
en_ambos = inv_materiales & formula_materiales

print(f"\n   📊 Resultados:")
print(f"      • En AMBOS:                    {len(en_ambos)}")
print(f"      • Solo en INVENTARIO:          {len(solo_en_inventario)}")
print(f"      • Solo en FORMULAS:            {len(solo_en_formulas)}")
print(f"      • TOTAL materiales únicos:     {len(inv_materiales | formula_materiales)}")

# 5. LISTA MAESTRA CONSOLIDADA
print("\n5️⃣  Generando LISTA MAESTRA CONSOLIDADA...")

materiales_consolidados = sorted(inv_materiales | formula_materiales)

sql_lines = []
for i, mat in enumerate(materiales_consolidados, 1):
    codigo = f"MPMP{i:05d}"
    mat_sql = mat.replace("'", "''")
    sql_lines.append(f"  ('{codigo}', '{mat_sql}', 'KG', TRUE)")

sql_content = f"""-- LISTA MAESTRA CONSOLIDADA
-- Inventario Real + Formulas Maestras
-- Total: {len(materiales_consolidados)} materiales

INSERT INTO materiales (codigo, nombre_inci, unidad, activo)
VALUES
""" + ",\n".join(sql_lines) + ";"

with open("LISTA_MAESTRA_CONSOLIDADA_FINAL.sql", "w", encoding="utf-8") as f:
    f.write(sql_content)

# 6. REPORTE DE DIFERENCIAS
print("\n6️⃣  Generando reportes...")

with open("CRUCE_INVENTARIO_vs_FORMULAS.txt", "w", encoding="utf-8") as f:
    f.write("ANÁLISIS CRUCE: INVENTARIO vs FORMULAS\n")
    f.write("=" * 80 + "\n\n")

    f.write(f"RESUMEN:\n")
    f.write(f"  • En AMBOS:               {len(en_ambos)}\n")
    f.write(f"  • Solo INVENTARIO:        {len(solo_en_inventario)}\n")
    f.write(f"  • Solo FORMULAS:          {len(solo_en_formulas)}\n")
    f.write(f"  • TOTAL:                  {len(materiales_consolidados)}\n\n")

    if solo_en_inventario:
        f.write(f"\n⚠️  SOLO EN INVENTARIO ({len(solo_en_inventario)}):\n")
        for mat in sorted(solo_en_inventario):
            f.write(f"  • {mat}\n")

    if solo_en_formulas:
        f.write(f"\n⚠️  SOLO EN FORMULAS ({len(solo_en_formulas)}):\n")
        for mat in sorted(solo_en_formulas):
            f.write(f"  • {mat}\n")

# 7. LISTA CONSOLIDADA LEGIBLE
with open("LISTA_MAESTRA_CONSOLIDADA.txt", "w", encoding="utf-8") as f:
    f.write("ID    | CÓDIGO     | NOMBRE\n")
    f.write("=" * 80 + "\n")
    for i, mat in enumerate(materiales_consolidados, 1):
        f.write(f"{i:5d} | MPMP{i:05d} | {mat}\n")

print(f"\n✅ ARCHIVOS GENERADOS:")
print(f"   • LISTA_MAESTRA_CONSOLIDADA_FINAL.sql")
print(f"   • LISTA_MAESTRA_CONSOLIDADA.txt")
print(f"   • CRUCE_INVENTARIO_vs_FORMULAS.txt")
print(f"\n📊 TOTAL MATERIALES: {len(materiales_consolidados)}")
print("=" * 80)
