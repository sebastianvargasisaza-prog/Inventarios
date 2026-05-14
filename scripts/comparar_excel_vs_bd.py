"""Compara FORMULAS_MAESTRO_v2_1.xlsx vs formula_items en BD.

Detecta:
- Items en BD que NO están en Excel (residuos de fórmulas viejas)
- Items en Excel que NO están en BD (faltantes de import)
- Discrepancias de gramos
- Items extra: 'AGUA' agregada por mig 126 NO se considera residuo

Uso:
  python scripts/comparar_excel_vs_bd.py [path_db]
"""
import sys
import os
import sqlite3
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'api'))

EXCEL_PATH = r"C:\Users\sebas\Downloads\FORMULAS_MAESTRO_v2_1 (2).xlsx"
DB_PATH = sys.argv[1] if len(sys.argv) > 1 else os.environ.get('DB_PATH', '/tmp/test_compare.db')

# Mapeo de SHEET name → producto_nombre canónico en BD (de mig 121)
SHEET_TO_BD = {
    'Emulsión Hidratante B3 BHA': 'EMULSION HIDRATANTE  B3+BHA',
    'Esencia Centella Asiática': 'ESENCIA DE CENTELLA ASIATICA',
    'Limpiador Facial BHA 2%': 'LIMPIADOR FACIAL BHA 2%',
    'Limpiador Iluminador Ácido Kóji': 'LIMPIADOR ILUMINADOR ACIDO KOJICO',
    'Mascarilla Hidratante': 'MASCARILLA HIDRATANTE',
    'Suero Antioxidante Renova C': 'SUERO ANTIOXIDANTE RENOVA C10',
    'Suero Vitamina C': 'SUERO DE VITAMINA C+ FORMULA NUEVA',
    'Suero Hidratante AH 1.5%': 'SUERO HIDRATANTE AH 1.5%',
    'Suero Iluminador TRX': 'SUERO ILUMINADOR TRX',
    'Suero Multipéptidos': 'SUERO MULTIPEPTIDOS',
    'Suero Niacinamida 5%': 'SUERO DE NIACINAMIDA 5% FORMULA NUEVA',
    'Suero Exfoliante Nova PHA': 'SUERO EXFOLIANTE NOVA PHA',
    'AZ Híbrid Clear': 'AZ HIBRID CLEAR',
    'Contorno de Cafeína': 'CONTORNO DE CAFEINA',
    'Contorno de Ojos Multipéptidos': 'CONTORNO DE OJOS MULTIPEPTIDOS',
    'Contorno de Ojos Retinaldehído': 'CONTORNO DE OJOS RETINALDEHIDO 0.05%',
    'Crema Corporal Renova Body': 'CREMA CORPORAL RENOVA BODY',
    'Limpiador Hidratante': 'LIMPIADOR FACIAL HIDRATANTE',
    'Suero Triactive Retinoid + NAD': 'SUERO TRIACTIVE RETINOID NAD',
    'Suero Exfoliante BHA 2%': 'Suero Exfoliante BHA 2%',
    'Gel Hidratante': 'GEL HIDRATANTE',
    'Booster Tensor': 'BOOSTER TENSOR',
    'Blush Balm': 'BLUSH BALM',
    'Emulsión Limpiadora': 'EMULSION LIMPIADORA',
    'HydraPeptide': 'HYDRAPEPTIDE',
    'Emulsión Hidratante Iluminadora': 'EMULSION HIDRATANTE ILUMINADORA',
    'Lip Sérum Voluminizador': 'LIP SERUM VOLUMINIZADOR PEPTIDOS',
    'Hydra-Balance': 'HYDRA BALANCE',
}


def main():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 1) Leer Excel · construir {producto_bd: {codigo_mp: g_por_kg}}
    excel_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == 'RESUMEN':
            continue
        producto_bd = SHEET_TO_BD.get(sheet_name)
        if not producto_bd:
            print(f"⚠ Hoja sin mapeo: {sheet_name}")
            continue
        ws = wb[sheet_name]
        items = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or row[0] is None or str(row[0]).strip() in ('', 'TOTAL'):
                continue
            cod = str(row[3]).strip() if row[3] else ''
            try:
                g_por_kg = float(row[5]) if row[5] is not None else 0
            except Exception:
                g_por_kg = 0
            if cod and cod.startswith('MP') and g_por_kg > 0:
                items[cod] = g_por_kg
        excel_data[producto_bd] = items

    # 2) Leer BD
    if not os.path.exists(DB_PATH):
        print(f"❌ DB no existe: {DB_PATH}")
        print("   Generala primero con: python -c \"from api.database import init_db; init_db()\" (con DB_PATH apuntando a una nueva)")
        return
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    bd_data = {}
    for r in c.execute("""
        SELECT producto_nombre, material_id, cantidad_g_por_lote, material_nombre
        FROM formula_items
        WHERE material_id IS NOT NULL AND TRIM(material_id) != ''
    """).fetchall():
        prod = r[0]
        bd_data.setdefault(prod, {})[r[1]] = (r[2], r[3])

    # 3) Comparar producto por producto
    print(f"\n{'=' * 80}")
    print(f"COMPARACIÓN EXCEL vs BD")
    print(f"{'=' * 80}\n")

    total_residuos = 0
    total_faltantes = 0
    total_discrepancias = 0

    for producto_bd in sorted(excel_data.keys()):
        excel_items = excel_data[producto_bd]
        bd_items = bd_data.get(producto_bd, {})
        if not bd_items:
            print(f"❌ {producto_bd}: NO existe en BD (Excel tiene {len(excel_items)} items)")
            continue

        residuos = []  # En BD pero no en Excel
        faltantes = []  # En Excel pero no en BD
        discrepancias = []  # Existen en ambos pero gramos diferentes

        for cod_bd, (cant_bd, nombre_bd) in bd_items.items():
            if cod_bd == 'MPAGUALI01':
                continue  # Agua agregada por mig 126 · esperado, no es residuo
            if cod_bd not in excel_items:
                residuos.append((cod_bd, cant_bd, nombre_bd))
            else:
                # Comparar gramos (escalado al lote BD)
                # Excel está en g/1kg · BD en g/lote total
                # Verificar via: cant_bd / lote_size_kg ≈ excel_g_por_kg
                lote_kg = None
                row_lote = c.execute(
                    "SELECT lote_size_kg FROM formula_headers WHERE producto_nombre = ?",
                    (producto_bd,)
                ).fetchone()
                if row_lote:
                    lote_kg = float(row_lote[0] or 0)
                if lote_kg and lote_kg > 0:
                    esperado_g = excel_items[cod_bd] * lote_kg
                    if abs(esperado_g - (cant_bd or 0)) > 1:  # tolerancia 1g
                        discrepancias.append((cod_bd, cant_bd, esperado_g, nombre_bd))

        for cod_ex, g_kg_ex in excel_items.items():
            if cod_ex not in bd_items:
                faltantes.append((cod_ex, g_kg_ex))

        if residuos or faltantes or discrepancias:
            print(f"\n📋 {producto_bd}")
            if residuos:
                total_residuos += len(residuos)
                print(f"  🔴 RESIDUOS (en BD pero NO en Excel · {len(residuos)}):")
                for cod, cant, nombre in residuos:
                    print(f"    {cod} · {nombre} · {cant}g")
            if faltantes:
                total_faltantes += len(faltantes)
                print(f"  🟠 FALTANTES (en Excel pero NO en BD · {len(faltantes)}):")
                for cod, g_kg in faltantes:
                    print(f"    {cod} · {g_kg}g/1kg")
            if discrepancias:
                total_discrepancias += len(discrepancias)
                print(f"  🟡 DISCREPANCIAS de cantidad ({len(discrepancias)}):")
                for cod, cant_bd, esperado, nombre in discrepancias:
                    print(f"    {cod} · {nombre} · BD={cant_bd}g · esperado={esperado:.1f}g")
        else:
            print(f"✅ {producto_bd}: OK ({len(bd_items)} items match Excel)")

    print(f"\n{'=' * 80}")
    print(f"RESUMEN GLOBAL:")
    print(f"  🔴 Total residuos (basura BD): {total_residuos}")
    print(f"  🟠 Total faltantes (no importados): {total_faltantes}")
    print(f"  🟡 Total discrepancias: {total_discrepancias}")
    print(f"{'=' * 80}\n")


if __name__ == '__main__':
    main()
