"""Genera las SQL statements de la mig 121 desde el Excel
FORMULAS_MAESTRO_v2_1 Alejandro mayo-2026.

Output: prints las statements como lista Python · copy-paste a database.py
"""
import openpyxl
import re
import unicodedata
import sys

EXCEL = r'C:\Users\sebas\Downloads\FORMULAS_MAESTRO_v2_1 (2).xlsx'

# Mapeo sheet → producto_nombre canónico en BD (de formula_headers existente)
# Para productos NUEVOS · usar el nombre del Excel directamente
SHEET_TO_BD = {
    'Emulsión Hidratante B3 BHA': 'EMULSION HIDRATANTE  B3+BHA',  # nota: 2 espacios + signo
    'Esencia Centella Asiática': 'ESENCIA DE CENTELLA ASIATICA',
    'Limpiador Facial BHA 2%': 'LIMPIADOR FACIAL BHA 2%',
    'Limpiador Iluminador Ácido Kóji': 'LIMPIADOR ILUMINADOR ACIDO KOJICO',  # sheet truncado a 31 chars
    'Mascarilla Hidratante': 'MASCARILLA HIDRATANTE',
    'Suero Antioxidante Renova C': 'SUERO ANTIOXIDANTE RENOVA C10',
    'Suero Vitamina C': 'SUERO DE VITAMINA C+ FORMULA NUEVA',
    'Suero Hidratante AH 1.5%': 'SUERO HIDRATANTE AH 1.5%',
    'Suero Iluminador TRX': 'SUERO ILUMINADOR TRX',
    'Suero Multipéptidos': 'SUERO MULTIPEPTIDOS',
    'Suero Niacinamida 5%': 'SUERO DE NIACINAMIDA 5% FORMULA NUEVA',
    'Suero Exfoliante Nova PHA': 'SUERO EXFOLIANTE NOVA PHA',
    'AZ Híbrid Clear': 'AZ HIBRID CLEAR',
    'Contorno de Cafeína': 'CONTORNO DE CAFEINA',
    'Contorno de Ojos Multipéptidos': 'CONTORNO DE OJOS MULTIPEPTIDOS',
    'Contorno de Ojos Retinaldehído': 'CONTORNO DE OJOS RETINALDEHIDO 0.05%',  # sheet truncado
    'Contorno de Ojos Retinaldehído ': 'CONTORNO DE OJOS RETINALDEHIDO 0.05%',  # con espacio final
    'Crema Corporal Renova Body': 'CREMA CORPORAL RENOVA BODY',
    'Limpiador Hidratante': 'LIMPIADOR FACIAL HIDRATANTE',
    'Suero Triactive Retinoid + NAD': 'SUERO TRIACTIVE RETINOID NAD',  # nuevo
    'Suero Exfoliante BHA 2%': 'Suero Exfoliante BHA 2%',  # case-mixed en BD
    'Gel Hidratante': 'GEL HIDRATANTE',
    'Booster Tensor': 'BOOSTER TENSOR',  # nuevo
    'Blush Balm': 'Blush Balm',  # case-mixed en BD
    'Emulsión Limpiadora': 'EMULSION LIMPIADORA',
    'HydraPeptide': 'HYDRAPEPTIDE',  # nuevo
    'Emulsión Hidratante Iluminadora': 'EMULSION HIDRATANTE ILUMINADORA',  # pendiente
    'Lip Sérum Voluminizador': 'LIP SERUM VOLUMINIZADOR PEPTIDOS',  # mapea al de péptidos
    'Hydra-Balance': 'HYDRA BALANCE',  # nuevo
}

# Lote_size_kg del RESUMEN · keys = sheet name del Excel
LOTE_KG = {
    'Emulsión Hidratante B3 BHA': 40,
    'Esencia Centella Asiática': 40,
    'Limpiador Facial BHA 2%': 120,
    'Limpiador Iluminador Ácido Kójico': 90,
    'Limpiador Iluminador Ácido Kóji': 90,  # truncado Excel
    'Mascarilla Hidratante': 10,
    'Suero Antioxidante Renova C': 20,
    'Suero Vitamina C': 20,
    'Suero Hidratante AH 1.5%': 90,
    'Suero Iluminador TRX': 100,
    'Suero Multipéptidos': 35,
    'Suero Niacinamida 5%': 100,
    'Suero Exfoliante Nova PHA': 14,
    'AZ Híbrid Clear': 33,
    'Contorno de Cafeína': 10,
    'Contorno de Ojos Multipéptidos': 15,
    'Contorno de Ojos Retinaldehído 0.05%': 13,
    'Contorno de Ojos Retinaldehído': 13,  # truncado
    'Contorno de Ojos Retinaldehído ': 13,  # con espacio
    'Crema Corporal Renova Body': 60,
    'Limpiador Hidratante': 80,
    'Suero Triactive Retinoid + NAD': 0.2,  # piloto · pendiente confirmar lote real
    'Suero Exfoliante BHA 2%': 0.1,  # piloto
    'Gel Hidratante': 50,
    'Booster Tensor': 1,
    'Blush Balm': 0,  # variable según tono · skip
    'Emulsión Limpiadora': 2,
    'HydraPeptide': 2,
    'Hydra-Balance': 0,  # variable · sin cantidades
    'Lip Sérum Voluminizador': 0,  # variable PIB
    'Emulsión Hidratante Iluminadora': 2,  # pendiente
}

# Productos PENDIENTES · NO importar items
# Sebastián 13-may-2026: "las otras dos las sacamos y las dejamos
# pendientes para que Alejandro no las envíe".
# Lip Sérum sí se importa (foto confirma fórmula del Excel hoja).
PENDIENTES = {
    'Emulsión Hidratante Iluminadora',  # v8 en proceso de cambios
    'Hydra-Balance',                     # sin cantidades · variable
}

# Productos NUEVOS · necesitan INSERT a formula_headers ANTES del INSERT a items
PRODUCTOS_NUEVOS = {
    'SUERO TRIACTIVE RETINOID NAD',
    'BOOSTER TENSOR',
    'HYDRAPEPTIDE',
    'HYDRA BALANCE',
}


def sql_escape(s):
    """Escape SQL single quotes."""
    if s is None:
        return "''"
    return "'" + str(s).replace("'", "''") + "'"


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    stmts = []
    productos_procesados = []

    # Paso 0: garantizar que los 146 codigos del Excel existan en maestro_mps
    # con su nombre INCI + comercial. INSERT OR IGNORE preserva existentes
    # (no pisa nombres que prod ya tiene). UPDATE activo=1 para los inactivos.
    import json
    info_path = r'C:\Users\sebas\Downloads\Claude\Inventarios\scripts\excel_mp_codigos.json'
    with open(info_path, encoding='utf-8') as f:
        excel_info = json.load(f)
    valores = ',\n  '.join(
        f"({sql_escape(cod)}, {sql_escape(excel_info[cod].get('comercial', ''))}, "
        f"{sql_escape(excel_info[cod].get('inci', ''))}, 1)"
        for cod in sorted(excel_info.keys())
    )
    stmts.append(
        "INSERT OR IGNORE INTO maestro_mps (codigo_mp, nombre_comercial, nombre_inci, activo) VALUES\n  "
        + valores
    )
    stmts.append(
        f"UPDATE maestro_mps SET activo = 1 WHERE codigo_mp IN ("
        + ",".join(sql_escape(c) for c in sorted(excel_info.keys()))
        + ")"
    )

    for sheet_name in wb.sheetnames:
        if sheet_name == 'RESUMEN':
            continue

        producto_bd = SHEET_TO_BD.get(sheet_name)
        if not producto_bd:
            print(f"# WARN sheet sin mapeo: {sheet_name}", file=sys.stderr)
            continue
        if sheet_name in PENDIENTES:
            print(f"# SKIP pendiente: {sheet_name}", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        lote_kg = LOTE_KG.get(sheet_name, 0)
        items = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or row[0] is None or str(row[0]).strip() in ('', 'TOTAL', '#'):
                continue
            try:
                num = int(str(row[0]).strip())
            except (ValueError, TypeError):
                continue
            nom_inci = str(row[1] or '').strip() if len(row) > 1 else ''
            nom_com = str(row[2] or '').strip() if len(row) > 2 else ''
            cod = str(row[3] or '').strip() if len(row) > 3 else ''
            if not re.match(r'^MP\d{5}$', cod):
                continue
            # row[4] = % fórmula (decimal · 0.01 = 1%)
            try:
                pct_decimal = float(row[4] or 0)
            except (ValueError, TypeError):
                pct_decimal = 0.0
            if pct_decimal <= 0:
                continue
            # Convertir a porcentaje BD (1.0 = 1%)
            pct_bd = pct_decimal * 100
            # cantidad_g_por_lote = pct × 1000 × lote_kg (lote en kg → gramos)
            if lote_kg > 0:
                cant_g = pct_decimal * 1000 * lote_kg
            else:
                cant_g = 0.0
            items.append({
                'codigo': cod,
                'nombre': nom_com or nom_inci,
                'pct': round(pct_bd, 6),
                'cant_g': round(cant_g, 4),
            })

        if not items:
            print(f"# SKIP sin items: {sheet_name}", file=sys.stderr)
            continue

        # Reset items previos
        stmts.append(
            f"DELETE FROM formula_items WHERE producto_nombre = {sql_escape(producto_bd)}"
        )

        # Crear/actualizar header
        if producto_bd in PRODUCTOS_NUEVOS:
            # Para productos nuevos · INSERT OR IGNORE (no perder activo=1 si ya existe)
            stmts.append(
                f"INSERT OR IGNORE INTO formula_headers (producto_nombre, unidad_base_g, lote_size_kg, activo, descripcion) "
                f"VALUES ({sql_escape(producto_bd)}, {lote_kg * 1000}, {lote_kg}, 1, "
                f"{sql_escape(f'Importado del Excel Alejandro mayo-2026 · {sheet_name}')})"
            )

        # Actualizar lote_size_kg + unidad_base_g + reactivar
        if lote_kg > 0:
            stmts.append(
                f"UPDATE formula_headers SET lote_size_kg = {lote_kg}, "
                f"unidad_base_g = {lote_kg * 1000}, activo = 1 "
                f"WHERE producto_nombre = {sql_escape(producto_bd)}"
            )
        else:
            stmts.append(
                f"UPDATE formula_headers SET activo = 1 "
                f"WHERE producto_nombre = {sql_escape(producto_bd)}"
            )

        # Insert items en bulk · 1 statement con múltiples VALUES
        if items:
            valores = ',\n  '.join(
                f"({sql_escape(producto_bd)}, {sql_escape(it['codigo'])}, "
                f"{sql_escape(it['nombre'])}, {it['pct']}, {it['cant_g']})"
                for it in items
            )
            stmts.append(
                "INSERT INTO formula_items (producto_nombre, material_id, "
                "material_nombre, porcentaje, cantidad_g_por_lote) VALUES\n  "
                + valores
            )

        productos_procesados.append((producto_bd, len(items), lote_kg))

    # Paso final · auto-seed MBR draft para productos nuevos
    # (mismo patrón que mig 115 · idempotente con NOT EXISTS)
    for nuevo in sorted(PRODUCTOS_NUEVOS):
        stmts.append(
            f"INSERT OR IGNORE INTO mbr_templates "
            f"(producto_nombre, version, estado, titulo, descripcion, "
            f"lote_size_g, tiempo_total_estimado_min, creado_por) "
            f"SELECT {sql_escape(nuevo)}, 1, 'draft', "
            f"{sql_escape(nuevo + ' · MBR v1 (auto-seed mig 121)')}, "
            f"'Procedimiento auto-generado desde formula_headers · "
            f"PENDIENTE Calidad debe ajustar pasos antes de aprobar.', "
            f"COALESCE(fh.unidad_base_g, 1000.0), 270, 'system-seed' "
            f"FROM formula_headers fh WHERE fh.producto_nombre = {sql_escape(nuevo)} "
            f"AND NOT EXISTS (SELECT 1 FROM mbr_templates m WHERE m.producto_nombre = {sql_escape(nuevo)})"
        )
        # 3 pasos típicos (idempotente vía NOT EXISTS)
        for orden, fase, desc, tipo, eq in [
            (1, 'dispensacion', 'Pesar y dispensar las MPs según fórmula maestra', 'dispensacion', 'BAL01,DISP'),
            (2, 'fabricacion', 'Fabricar el producto · PENDIENTE Calidad detalle parámetros', 'mezclado', 'TQ01'),
            (3, 'envasado', 'Envasar y etiquetar · QC firma liberación', 'envasado', 'ENV1'),
        ]:
            stmts.append(
                f"INSERT INTO mbr_pasos (mbr_template_id, orden, fase, descripcion, "
                f"tipo_paso, equipo_requerido, tiempo_estimado_min, requiere_e_sign, "
                f"requiere_qc, notas) "
                f"SELECT m.id, {orden}, {sql_escape(fase)}, {sql_escape(desc)}, "
                f"{sql_escape(tipo)}, {sql_escape(eq)}, 60, "
                f"{1 if orden == 1 else 0}, {1 if orden == 3 else 0}, '' "
                f"FROM mbr_templates m WHERE m.producto_nombre = {sql_escape(nuevo)} "
                f"AND m.creado_por = 'system-seed' "
                f"AND NOT EXISTS (SELECT 1 FROM mbr_pasos p WHERE p.mbr_template_id = m.id AND p.orden = {orden})"
            )

    # Output Python literal list
    print("# Generado por scripts/generate_mig_121_formulas.py · 13-may-2026")
    print(f"# {len(productos_procesados)} productos · {sum(p[1] for p in productos_procesados)} items totales")
    print("# Productos:")
    for prod, n, lote in productos_procesados:
        print(f"#   {prod}: {n} items · lote {lote}kg")
    print()
    print("# Statements:")
    for s in stmts:
        # Output as Python triple-quoted string per stmt
        print(f'    """{s}""",')


if __name__ == '__main__':
    main()
