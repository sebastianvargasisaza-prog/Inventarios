"""Regenera mig 127 desde el Excel maestro · re-import COMPLETO.

Hace lo siguiente:
1. Lee TODAS las hojas del Excel FORMULAS_MAESTRO_v2_1.xlsx
2. Para cada producto, genera SQL idempotente:
   - INSERT OR IGNORE maestro_mps para cada material_id usado
   - INSERT OR IGNORE formula_headers (lote_size_kg + nombre canónico)
   - DELETE formula_items WHERE producto_nombre = X (limpia versión anterior)
   - INSERT formula_items con datos exactos del Excel
3. Escribe a api/mig_127_data.py · módulo importable desde database.py
"""
import openpyxl

EXCEL = r"C:\Users\sebas\Downloads\FORMULAS_MAESTRO_v2_1 (2).xlsx"
OUTPUT = "api/mig_127_data.py"

# Sheet name → producto_nombre canónico en BD
SHEET_TO_BD = {
    'Emulsión Hidratante B3 BHA': 'EMULSION HIDRATANTE  B3+BHA',
    'Esencia Centella Asiática': 'ESENCIA DE CENTELLA ASIATICA',
    'Limpiador Facial BHA 2%': 'LIMPIADOR FACIAL BHA 2%',
    'Limpiador Iluminador Ácido Kóji': 'LIMPIADOR ILUMINADOR ACIDO KOJICO',
    'Mascarilla Hidratante': 'MASCARILLA HIDRATANTE',
    'Suero Antioxidante Renova C': 'SUERO ANTIOXIDANTE RENOVA C10',
    'Suero Vitamina C': 'SUERO DE VITAMINA C+ FORMULA NUEVA',
    'Suero Hidratante AH 1.5%': 'SUERO HIDRATANTE AH 1.5%',
    'Suero Iluminador TRX': 'SUERO ILUMINADOR TRX',
    'Suero Multipéptidos': 'SUERO MULTIPEPTIDOS',
    'Suero Niacinamida 5%': 'SUERO DE NIACINAMIDA 5% FORMULA NUEVA',
    'Suero Exfoliante Nova PHA': 'SUERO EXFOLIANTE NOVA PHA',
    'AZ Híbrid Clear': 'AZ HIBRID CLEAR',
    'Contorno de Cafeína': 'CONTORNO DE CAFEINA',
    'Contorno de Ojos Multipéptidos': 'CONTORNO DE OJOS MULTIPEPTIDOS',
    'Contorno de Ojos Retinaldehído': 'CONTORNO DE OJOS RETINALDEHIDO 0.05%',
    'Contorno de Ojos Retinaldehído ': 'CONTORNO DE OJOS RETINALDEHIDO 0.05%',
    'Crema Corporal Renova Body': 'CREMA CORPORAL RENOVA BODY',
    'Limpiador Hidratante': 'LIMPIADOR FACIAL HIDRATANTE',
    'Suero Triactive Retinoid + NAD': 'SUERO TRIACTIVE RETINOID NAD',
    'Suero Exfoliante BHA 2%': 'Suero Exfoliante BHA 2%',
    'Gel Hidratante': 'GEL HIDRATANTE',
    'Booster Tensor': 'BOOSTER TENSOR',
    'Blush Balm': 'BLUSH BALM',
    'Emulsión Limpiadora': 'EMULSION LIMPIADORA',
    'HydraPeptide': 'HYDRAPEPTIDE',
    'Emulsión Hidratante Iluminadora': 'EMULSION HIDRATANTE ILUMINADORA',
    'Lip Sérum Voluminizador': 'LIP SERUM VOLUMINIZADOR PEPTIDOS',
    'Hydra-Balance': 'HYDRA BALANCE',
}

# Lote_size_kg por sheet (del RESUMEN del Excel)
LOTE_KG = {
    'Emulsión Hidratante B3 BHA': 40,
    'Esencia Centella Asiática': 40,
    'Limpiador Facial BHA 2%': 120,
    'Limpiador Iluminador Ácido Kóji': 90,
    'Mascarilla Hidratante': 10,
    'Suero Antioxidante Renova C': 20,
    'Suero Vitamina C': 20,
    'Suero Hidratante AH 1.5%': 90,
    'Suero Iluminador TRX': 100,
    'Suero Multipéptidos': 35,
    'Suero Niacinamida 5%': 100,
    'Suero Exfoliante Nova PHA': 14,
    'AZ Híbrid Clear': 33,
    'Contorno de Cafeína': 10,
    'Contorno de Ojos Multipéptidos': 15,
    'Contorno de Ojos Retinaldehído': 13,
    'Contorno de Ojos Retinaldehído ': 13,
    'Crema Corporal Renova Body': 60,
    'Limpiador Hidratante': 80,
    'Suero Triactive Retinoid + NAD': 0.2,
    'Suero Exfoliante BHA 2%': 0.1,
    'Gel Hidratante': 50,
    'Booster Tensor': 1,
    'Blush Balm': 1,  # asumimos 1kg para que la fórmula funcione
    'Emulsión Limpiadora': 2,
    'HydraPeptide': 2,
    'Emulsión Hidratante Iluminadora': 30,
    'Lip Sérum Voluminizador': 1,
    'Hydra-Balance': 50,  # estimado
}


def sql_quote(s):
    """Escapa string para SQL."""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    statements = []
    statements.append("-- mig 127 · re-import COMPLETO desde Excel mayo-2026")
    statements.append("-- Generado por scripts/generate_mig_127_reimport.py")

    # Recolectar todos los materiales únicos del Excel (cod → (nombre_inci, nombre_comercial))
    materiales_globales = {}
    formulas_data = {}  # producto_bd → (lote_kg, [(cod, nombre_inci, nombre_comercial, g_por_kg)])

    for sheet_name in wb.sheetnames:
        if sheet_name == 'RESUMEN':
            continue
        producto_bd = SHEET_TO_BD.get(sheet_name)
        if not producto_bd:
            print(f"⚠ Hoja sin mapeo: {sheet_name!r}")
            continue
        lote_kg = LOTE_KG.get(sheet_name)
        if lote_kg is None:
            print(f"⚠ Sin lote_kg para: {sheet_name!r}")
            continue
        ws = wb[sheet_name]
        items = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or row[0] is None or str(row[0]).strip() in ('', 'TOTAL'):
                continue
            nombre_inci = str(row[1]).strip() if row[1] else ''
            nombre_com = str(row[2]).strip() if row[2] else ''
            cod = str(row[3]).strip() if row[3] else ''
            try:
                g_por_kg = float(row[5]) if row[5] is not None else 0
            except Exception:
                g_por_kg = 0
            if cod and cod.startswith('MP') and g_por_kg > 0:
                items.append((cod, nombre_inci, nombre_com, g_por_kg))
                # Guardar info global
                if cod not in materiales_globales:
                    materiales_globales[cod] = (nombre_inci, nombre_com)
        if items:
            formulas_data[producto_bd] = (lote_kg, items)

    print(f"\nTotal productos: {len(formulas_data)}")
    print(f"Total materiales únicos: {len(materiales_globales)}")
    total_items = sum(len(items) for _, items in formulas_data.values())
    print(f"Total items: {total_items}\n")

    # ── Statements ──
    # Paso 1 · Asegurar maestro_mps tenga TODOS los códigos del Excel
    statements.append("\n-- ── Paso 1 · maestro_mps · upsert de TODOS los códigos Excel ──")
    for cod, (inci, com) in sorted(materiales_globales.items()):
        statements.append(
            f"INSERT OR IGNORE INTO maestro_mps "
            f"(codigo_mp, nombre_comercial, nombre_inci, activo, stock_minimo, proveedor) "
            f"VALUES ({sql_quote(cod)}, {sql_quote(com or inci)}, {sql_quote(inci)}, 1, 0, '')"
        )
        statements.append(
            f"UPDATE maestro_mps SET activo = 1 "
            f"WHERE codigo_mp = {sql_quote(cod)} AND COALESCE(activo, 0) = 0"
        )

    # Paso 2 · formula_headers · upsert
    statements.append("\n-- ── Paso 2 · formula_headers · upsert lote_size_kg ──")
    for producto, (lote_kg, _) in sorted(formulas_data.items()):
        statements.append(
            f"INSERT OR IGNORE INTO formula_headers "
            f"(producto_nombre, lote_size_kg, unidad_base_g, activo, descripcion) "
            f"VALUES ({sql_quote(producto)}, {lote_kg}, {lote_kg * 1000}, 1, "
            f"'Re-import mig 127 · Excel mayo-2026')"
        )
        statements.append(
            f"UPDATE formula_headers SET "
            f"lote_size_kg = {lote_kg}, unidad_base_g = {lote_kg * 1000}, activo = 1 "
            f"WHERE producto_nombre = {sql_quote(producto)}"
        )

    # Paso 3 · Borrar formula_items VIEJOS de productos que vamos a re-importar
    statements.append("\n-- ── Paso 3 · borrar items VIEJOS (limpia residuos) ──")
    for producto in sorted(formulas_data.keys()):
        statements.append(
            f"DELETE FROM formula_items WHERE producto_nombre = {sql_quote(producto)}"
        )

    # Paso 4 · Insertar items nuevos con cantidad_g_por_lote correcta
    statements.append("\n-- ── Paso 4 · insertar items del Excel (g/kg × lote_kg) ──")
    for producto, (lote_kg, items) in sorted(formulas_data.items()):
        for cod, inci, com, g_por_kg in items:
            cantidad_g = round(g_por_kg * lote_kg, 4)
            porcentaje = round(g_por_kg / 10.0, 6)  # g/kg → %
            statements.append(
                f"INSERT INTO formula_items "
                f"(producto_nombre, material_id, material_nombre, porcentaje, cantidad_g_por_lote) "
                f"VALUES ({sql_quote(producto)}, {sql_quote(cod)}, {sql_quote(com or inci)}, "
                f"{porcentaje}, {cantidad_g})"
            )

    # Paso 5 · Agregar AGUA q.s.p. (mig 126 lógica · idempotente)
    statements.append("\n-- ── Paso 5 · agregar AGUA q.s.p. (lote - suma items) ──")
    statements.append(
        "INSERT OR IGNORE INTO maestro_mps "
        "(codigo_mp, nombre_comercial, nombre_inci, activo, stock_minimo, proveedor) "
        "VALUES ('MPAGUALI01', 'Agua Desionizada', 'AQUA', 1, 0, 'Planta')"
    )
    statements.append(
        "UPDATE maestro_mps SET activo = 1 "
        "WHERE codigo_mp = 'MPAGUALI01' AND COALESCE(activo, 0) = 0"
    )
    for producto, (lote_kg, items) in sorted(formulas_data.items()):
        suma_g = sum(g * lote_kg for _, _, _, g in items)
        agua_g = round(lote_kg * 1000 - suma_g, 4)
        if agua_g > 0.1:  # solo si realmente falta agua
            pct_agua = round((agua_g / (lote_kg * 1000)) * 100, 4)
            statements.append(
                f"INSERT INTO formula_items "
                f"(producto_nombre, material_id, material_nombre, porcentaje, cantidad_g_por_lote) "
                f"VALUES ({sql_quote(producto)}, 'MPAGUALI01', 'AGUA DESIONIZADA', "
                f"{pct_agua}, {agua_g})"
            )

    # Paso 6 · Auto-seed MBR para productos nuevos (BLUSH BALM, HYDRA BALANCE,
    # LIP SERUM VOLUMINIZADOR PEPTIDOS · re-correr mig 115)
    statements.append("\n-- ── Paso 6 · auto-seed MBR para productos nuevos (idempotente) ──")
    statements.append(
        """INSERT OR IGNORE INTO mbr_templates
             (producto_nombre, version, estado, titulo, descripcion,
              lote_size_g, tiempo_total_estimado_min, creado_por)
           SELECT
             fh.producto_nombre, 1, 'draft',
             fh.producto_nombre || ' · MBR v1 (auto-seed mig 127)',
             'Auto-seed mig 127 · re-import Excel · PENDIENTE Calidad ajusta pasos.',
             COALESCE(fh.unidad_base_g, 1000.0), 270, 'system-seed'
           FROM formula_headers fh
           WHERE NOT EXISTS (
             SELECT 1 FROM mbr_templates m
             WHERE m.producto_nombre = fh.producto_nombre
           )"""
    )
    # Pasos para los nuevos MBR (mismos 3 pasos que mig 115)
    statements.append(
        """INSERT INTO mbr_pasos
             (mbr_template_id, orden, fase, descripcion, tipo_paso,
              equipo_requerido, tiempo_estimado_min, requiere_e_sign,
              requiere_qc, notas)
           SELECT m.id, 1, 'dispensacion',
                  'Pesar y dispensar las MPs según fórmula maestra',
                  'dispensacion', 'BAL01,DISP', 60, 1, 0,
                  'Verificar lote y vencimiento de cada MP.'
           FROM mbr_templates m
           WHERE m.creado_por = 'system-seed' AND m.estado = 'draft'
             AND NOT EXISTS (SELECT 1 FROM mbr_pasos p WHERE p.mbr_template_id = m.id)"""
    )
    statements.append(
        """INSERT INTO mbr_pasos
             (mbr_template_id, orden, fase, descripcion, tipo_paso,
              equipo_requerido, tiempo_estimado_min, requiere_e_sign,
              requiere_qc, notas)
           SELECT m.id, 2, 'fabricacion',
                  'Fabricar el producto siguiendo procedimiento aprobado',
                  'mezclado', 'TQ01', 120, 0, 0, ''
           FROM mbr_templates m
           WHERE m.creado_por = 'system-seed' AND m.estado = 'draft'
             AND (SELECT COUNT(*) FROM mbr_pasos p WHERE p.mbr_template_id = m.id) = 1"""
    )
    statements.append(
        """INSERT INTO mbr_pasos
             (mbr_template_id, orden, fase, descripcion, tipo_paso,
              equipo_requerido, tiempo_estimado_min, requiere_e_sign,
              requiere_qc, notas)
           SELECT m.id, 3, 'envasado',
                  'Envasar producto y etiquetar · QC firma liberación',
                  'envasado', 'ENV1', 90, 0, 1,
                  'Verificar etiquetado, cierre, hermeticidad.'
           FROM mbr_templates m
           WHERE m.creado_por = 'system-seed' AND m.estado = 'draft'
             AND (SELECT COUNT(*) FROM mbr_pasos p WHERE p.mbr_template_id = m.id) = 2"""
    )

    # Escribir el archivo Python
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write('"""Datos para mig 127 · re-import COMPLETO Excel mayo-2026.\n')
        f.write('Generado automáticamente por scripts/generate_mig_127_reimport.py\n')
        f.write('NO editar a mano · regenerar el script si el Excel cambia.\n')
        f.write('"""\n\n')
        f.write('STATEMENTS = [\n')
        for stmt in statements:
            if stmt.startswith('--'):
                f.write(f'    # {stmt[2:].strip()}\n')
            else:
                # Escapar triple-quote y newlines
                stmt_safe = stmt.replace('"""', '\\"\\"\\"')
                f.write(f'    """{stmt_safe}""",\n')
        f.write(']\n')

    print(f"✓ Generado: {OUTPUT}")
    print(f"  Total statements: {sum(1 for s in statements if not s.startswith('--'))}")


if __name__ == '__main__':
    main()
