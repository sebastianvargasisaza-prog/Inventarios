"""
Sincroniza el banco de creadores del Excel con la tabla marketing_influencers.

Uso (local, una sola vez antes de promover):
    python scripts/sync_influencers_excel.py "C:/ruta/al/excel.xlsx"

Comportamiento:
  - UPSERT por nombre (case insensitive, espacios trim)
  - PRESERVA datos existentes (si ya hay banco/cuenta/cedula no los borra)
  - Solo actualiza campos vacíos en la DB con datos del Excel
  - Reporta diff: nuevos / actualizados / sin cambios
  - Modo --dry-run para ver el diff sin escribir
"""
import os
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


def _norm(s):
    return (s or "").strip().lower() if isinstance(s, str) else ""


def _str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def sync(excel_path, db_path, dry_run=False):
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel no encontrado: {excel_path}")
        return 1
    if not os.path.exists(db_path):
        print(f"ERROR: DB no encontrada: {db_path}")
        return 1

    wb = load_workbook(excel_path, data_only=True)
    if "cuentas de creadores" not in wb.sheetnames:
        print("ERROR: hoja 'cuentas de creadores' no existe")
        return 1

    ws = wb["cuentas de creadores"]
    creadores = []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [_norm(c) for c in row]
            continue
        if not row[0]:
            continue
        creadores.append({
            "nombre":     _str(row[0]),
            "banco":      _str(row[1]) if len(row) > 1 else "",
            "cuenta":     _str(row[2]) if len(row) > 2 else "",
            "tipo_cta":   _str(row[3]) if len(row) > 3 else "",
            "cedula":     _str(row[4]) if len(row) > 4 else "",
            "ciudad":     _str(row[5]) if len(row) > 5 else "",
            "user":       _str(row[6]) if len(row) > 6 else "",
            "tipo_creador": _str(row[7]) if len(row) > 7 else "",
        })

    print(f"Excel: {len(creadores)} creadores leídos")
    print()

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Verificar columnas existentes en marketing_influencers
    c.execute("PRAGMA table_info(marketing_influencers)")
    cols_db = {r["name"] for r in c.fetchall()}
    print(f"DB cols: {sorted(cols_db)}")
    print()

    # Mapeo Excel → DB. Agrega solo si la columna existe.
    EXCEL_TO_DB = {
        "banco":         "banco",
        "cuenta":        "cuenta_bancaria",
        "tipo_cta":      "tipo_cuenta",
        "cedula":        "cedula_nit",
        "ciudad":        "ciudad",
        "user":          "instagram",
        "tipo_creador":  "tipo",
    }

    nuevos, actualizados, sin_cambios = [], [], []

    for ex in creadores:
        nombre = ex["nombre"]
        # Buscar por nombre (LOWER + trim)
        c.execute(
            "SELECT * FROM marketing_influencers WHERE LOWER(TRIM(nombre)) = ? LIMIT 1",
            (nombre.lower(),)
        )
        existing = c.fetchone()

        if not existing:
            # INSERT
            cols_to_insert = ["nombre"]
            vals = [nombre]
            for k_ex, k_db in EXCEL_TO_DB.items():
                if k_db in cols_db and ex.get(k_ex):
                    cols_to_insert.append(k_db)
                    vals.append(ex[k_ex])
            placeholders = ",".join("?" * len(cols_to_insert))
            sql = f"INSERT INTO marketing_influencers ({','.join(cols_to_insert)}) VALUES ({placeholders})"
            if not dry_run:
                c.execute(sql, vals)
            nuevos.append(nombre)
            continue

        # UPDATE solo campos vacíos en DB
        sets, vals = [], []
        cambios_locales = []
        for k_ex, k_db in EXCEL_TO_DB.items():
            if k_db not in cols_db:
                continue
            new_val = ex.get(k_ex, "")
            old_val = existing[k_db] if k_db in existing.keys() else ""
            old_val = (old_val or "").strip() if isinstance(old_val, str) else ""
            if new_val and not old_val:
                sets.append(f"{k_db}=?")
                vals.append(new_val)
                cambios_locales.append(f"{k_db}=<vacío>→'{new_val[:30]}'")

        if sets:
            vals.append(existing["id"])
            if not dry_run:
                c.execute(
                    f"UPDATE marketing_influencers SET {','.join(sets)} WHERE id=?",
                    vals
                )
            actualizados.append((nombre, cambios_locales))
        else:
            sin_cambios.append(nombre)

    if not dry_run:
        conn.commit()

    conn.close()

    print("=" * 80)
    print(f"NUEVOS ({len(nuevos)}):")
    for n in nuevos:
        print(f"  + {n}")
    print()
    print(f"ACTUALIZADOS ({len(actualizados)}):")
    for n, cambios in actualizados:
        print(f"  ~ {n}")
        for c_ in cambios[:3]:
            print(f"      · {c_}")
    print()
    print(f"SIN CAMBIOS ({len(sin_cambios)}):")
    for n in sin_cambios:
        print(f"  = {n}")
    print()
    if dry_run:
        print("[DRY RUN — sin escribir a la DB]")
    else:
        print("[OK — DB actualizada]")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/sync_influencers_excel.py <excel.xlsx> [db.db] [--dry-run]")
        sys.exit(1)
    excel_path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith("--") else \
              os.environ.get("DB_PATH") or "/var/data/inventario.db"
    dry_run = "--dry-run" in sys.argv
    sys.exit(sync(excel_path, db_path, dry_run))
