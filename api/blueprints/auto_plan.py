"""
Auto-Plan Maestro — Sebastian + Alejandro (30-abr-2026)

"Tenemos algo maravilloso allí en planta... debe ser la herramienta más
avanzada del mundo, anclada y generada por Claude... usa toda tu capacidad
para que quede perfecta".

El motor que cada lunes a las 7am genera AUTOMÁTICAMENTE:
  1. Plan de producción próximas 8 semanas (solo L/M/V)
  2. Compras anticipadas de MP y envases
  3. Calendario de conteo cíclico (Ma/Ju)
  4. Emails ejecutivos a Sebastián, Alejandro, Catalina
  5. Agenda diaria por operario

Reglas (Sebastian, 30-abr-2026):
  - Vit C → cadencia 30d
  - Suero AH 1.5% → cadencia 90d, lote 90kg
  - Otros sueros/hidratantes/limpiadores → auto por umbral cobertura<30d
  - L/M/V producir, Ma/Ju acondicionar/envasar/conteo
  - MP mínimo 30d, ideal 60d
  - Envases mínimo 90d (China lead 180d)
  - Predicción agresiva exacta

Si dudas algo de MP → genera tarea "verificar existencia real".
"""
from flask import Blueprint, jsonify, request, session
from datetime import datetime, timedelta, date
from collections import defaultdict
import os
import sqlite3
import json
import logging
from database import get_db
from config import ADMIN_USERS

bp = Blueprint('auto_plan', __name__)
log = logging.getLogger('auto_plan')

# Días de la semana donde SE PRODUCE (lunes=0, martes=1, ...)
DIAS_PRODUCCION = (0, 2, 4)        # lunes, miércoles, viernes
DIAS_ACOND_CONTEO = (1, 3)         # martes, jueves
DIAS_NO_LABORAL = (5, 6)           # sábado, domingo

# ───────────────────────────────────────────────────────────────────────
# UTILIDADES
# ───────────────────────────────────────────────────────────────────────

def _next_dia_produccion(desde_fecha):
    """Devuelve la próxima fecha L/M/V desde una fecha base."""
    f = desde_fecha
    while f.weekday() not in DIAS_PRODUCCION:
        f += timedelta(days=1)
    return f


def _proximo_dia_acond(desde_fecha):
    """Próximo Ma o Ju (para conteo cíclico / acondicionamiento)."""
    f = desde_fecha
    while f.weekday() not in DIAS_ACOND_CONTEO:
        f += timedelta(days=1)
    return f


def _ventas_diarias_por_sku(c, sku, dias=60):
    """Devuelve [(fecha, unidades)] de ventas del SKU en los últimos N días.

    Detecta automáticamente la tabla disponible:
      1) ventas_diarias (sku, fecha, cantidad) — si existe
      2) animus_shopify_orders.sku_items (JSON parse) — caso real Espagiria
      3) ordenes_shopify_items legacy
    """
    # Estrategia 1: tabla agregada
    try:
        r = c.execute("""
            SELECT fecha, COALESCE(SUM(cantidad),0)
            FROM ventas_diarias WHERE sku=? AND fecha >= date('now','-' || ? || ' days')
            GROUP BY fecha ORDER BY fecha
        """, (sku, dias)).fetchall()
        if r:
            return [(row[0], float(row[1] or 0)) for row in r]
    except Exception:
        pass

    # Estrategia 2: animus_shopify_orders con sku_items JSON
    try:
        rows = c.execute("""
            SELECT date(creado_en), sku_items
            FROM animus_shopify_orders
            WHERE creado_en >= date('now','-' || ? || ' days')
              AND sku_items IS NOT NULL AND sku_items != ''
        """, (dias,)).fetchall()
        if rows:
            por_dia = {}
            for fecha, sku_items_json in rows:
                if not fecha or not sku_items_json:
                    continue
                try:
                    items = json.loads(sku_items_json) if isinstance(sku_items_json, str) else sku_items_json
                except Exception:
                    continue
                if not isinstance(items, list):
                    continue
                cantidad = 0
                for it in items:
                    sk = (it.get('sku') or it.get('SKU') or '').strip()
                    if sk == sku:
                        cantidad += float(it.get('cantidad') or it.get('quantity') or 0)
                if cantidad > 0:
                    por_dia[fecha] = por_dia.get(fecha, 0) + cantidad
            return sorted([(f, q) for f, q in por_dia.items()])
    except Exception:
        pass

    # Estrategia 3: ordenes_shopify_items legacy
    try:
        r = c.execute("""
            SELECT date(fecha), COALESCE(SUM(cantidad),0)
            FROM ordenes_shopify_items
            WHERE sku=? AND fecha >= date('now','-' || ? || ' days')
            GROUP BY date(fecha) ORDER BY 1
        """, (sku, dias)).fetchall()
        if r:
            return [(row[0], float(row[1] or 0)) for row in r]
    except Exception:
        pass

    return []


def _velocidad_y_tendencia(c, sku):
    """Velocidad de venta diaria + factor de tendencia (regresión 30d).

    factor_tendencia:
      1.0 = estable · >1 = subiendo · <1 = bajando
    """
    rows = _ventas_diarias_por_sku(c, sku, dias=30)

    if not rows:
        # Sin datos. Fallback: consumo desde stock_pt (movimientos negativos = ventas)
        try:
            r = c.execute("""
                SELECT COALESCE(SUM(unidades_disponible),0) FROM stock_pt WHERE sku=?
            """, (sku,)).fetchone()
            stock_actual = float(r[0] if r else 0)
            # Si tiene stock, asumimos rotación trimestral conservadora (1 unit/dia)
            return (max(0.5, stock_actual / 90.0), 1.0)
        except Exception:
            return (0.0, 1.0)

    n = len(rows)
    if n < 5:
        avg = sum(q for _, q in rows) / max(1, n)
        return (float(avg), 1.0)

    # Regresión lineal sobre días reales (no índices) para captar tendencia
    xs = list(range(n))
    ys = [q for _, q in rows]
    sum_x = sum(xs); sum_y = sum(ys)
    sum_xy = sum(x * y for x, y in zip(xs, ys))
    sum_xx = sum(x * x for x in xs)
    try:
        b = (n * sum_xy - sum_x * sum_y) / (n * sum_xx - sum_x * sum_x)
        a = (sum_y - b * sum_x) / n
    except ZeroDivisionError:
        b = 0
        a = sum_y / n
    velocidad_actual = max(0.0, a + b * (n - 1))
    velocidad_inicial = max(0.01, a + b * 0)
    factor = velocidad_actual / velocidad_inicial if velocidad_inicial > 0 else 1.0
    factor = max(0.3, min(factor, 3.0))
    # Velocidad observada simple (suma / días totales) como fallback
    obs = sum(ys) / 30.0
    velocidad_final = max(velocidad_actual, obs)
    return (velocidad_final, factor)


def _stock_actual_pt(c, producto):
    """Stock total PT actual sumando todos los SKUs del producto."""
    rows = c.execute(
        "SELECT sku FROM sku_producto_map WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND COALESCE(activo,1)=1",
        (producto,)
    ).fetchall()
    total = 0
    for (sku,) in rows:
        r = c.execute(
            "SELECT COALESCE(SUM(unidades_disponible),0) FROM stock_pt WHERE sku=?",
            (sku,)
        ).fetchone()
        if r:
            total += r[0] or 0
    return int(total)


def _velocidad_total_producto(c, producto):
    """Suma velocidades + tendencia agregada de todos los SKUs del producto."""
    rows = c.execute(
        "SELECT sku FROM sku_producto_map WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND COALESCE(activo,1)=1",
        (producto,)
    ).fetchall()
    vel_total = 0.0
    factores = []
    for (sku,) in rows:
        v, f = _velocidad_y_tendencia(c, sku)
        vel_total += v
        if v > 0:
            factores.append(f)
    factor_avg = (sum(factores) / len(factores)) if factores else 1.0
    return (vel_total, factor_avg)


def _alias_calendar_for(c, producto):
    """Devuelve el alias_calendar configurado del producto, o None."""
    r = c.execute(
        "SELECT alias_calendar FROM sku_planeacion_config WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    return r[0] if r and r[0] else None


def _ultima_produccion(c, producto):
    """Fecha de última producción del producto. Usa matcher robusto."""
    fechas = []
    # BD
    r = c.execute("""
        SELECT MAX(fecha_programada) FROM produccion_programada
        WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
          AND estado IN ('completado','en_proceso','pendiente')
    """, (producto,)).fetchone()
    if r and r[0]:
        try:
            fechas.append(datetime.strptime(r[0][:10], '%Y-%m-%d').date())
        except Exception:
            pass
    # Calendar con matcher robusto
    eventos = _calendar_events_cached()
    alias = _alias_calendar_for(c, producto)
    for ev in eventos:
        score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
        if score >= 60:  # threshold de confianza
            try:
                fechas.append(datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date())
            except Exception:
                pass
    return max(fechas) if fechas else None


def _historico_producciones_producto(c, producto, dias_atras=365):
    """Devuelve TODAS las fechas de producciones (BD + Calendar) con match robusto."""
    fechas = set()
    desde = (datetime.now().date() - timedelta(days=dias_atras)).isoformat()
    rows = c.execute("""
        SELECT fecha_programada FROM produccion_programada
        WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
          AND estado IN ('completado','en_proceso','pendiente')
          AND date(fecha_programada) >= ?
    """, (producto, desde)).fetchall()
    for r in rows:
        try:
            fechas.add(datetime.strptime(r[0][:10], '%Y-%m-%d').date())
        except Exception:
            pass
    # Calendar con matcher robusto
    eventos = _calendar_events_cached()
    alias = _alias_calendar_for(c, producto)
    for ev in eventos:
        score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
        if score >= 60:
            try:
                fechas.add(datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date())
            except Exception:
                pass
    return sorted(fechas)


_CAL_CACHE = {'ts': None, 'events': []}

def _calendar_events_cached(force_refresh=False):
    """Lee Google Calendar una vez por minuto y cachea."""
    now = datetime.now()
    if not force_refresh and _CAL_CACHE['ts'] and (now - _CAL_CACHE['ts']).total_seconds() < 60:
        return _CAL_CACHE['events']
    try:
        from blueprints.programacion import _fetch_calendar_events
        result = _fetch_calendar_events(days_ahead=180) or {}
        _CAL_CACHE['events'] = result.get('events') or []
        _CAL_CACHE['ts'] = now
    except Exception:
        _CAL_CACHE['events'] = []
        _CAL_CACHE['ts'] = now
    return _CAL_CACHE['events']


# ════════════════════════════════════════════════════════════════════════
# MATCHING ROBUSTO producto ↔ evento Calendar (zero-error-enterprise)
# ════════════════════════════════════════════════════════════════════════
import re as _re_match
import unicodedata as _ud

# Stopwords que NO deben usarse para matching (palabras genéricas)
_MATCH_STOPWORDS = {
    'SUERO', 'CREMA', 'GEL', 'EMULSION', 'LIMPIADOR', 'CONTORNO',
    'OJOS', 'FACIAL', 'CORPORAL', 'HIDRATANTE', 'ILUMINADOR',
    'ANTIOXIDANTE', 'EXFOLIANTE', 'MASCARILLA', 'ESENCIA', 'NUEVA',
    'FORMULA', 'PARA', 'CON', 'DE', 'LA', 'EL', 'BHA', 'AHA',
}

def _normalizar(s):
    """Normaliza string: quita tildes, mayúsculas, trim, sin caracteres raros."""
    if not s:
        return ''
    s = _ud.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not _ud.combining(c))
    return s.upper().strip()


def _palabras_clave_unicas(producto_nombre):
    """Extrae palabras CLAVE únicas del producto (no stopwords, >2 chars)."""
    norm = _normalizar(producto_nombre)
    palabras = _re_match.findall(r'[A-Z0-9+]+', norm)
    return [p for p in palabras if len(p) >= 2 and p not in _MATCH_STOPWORDS]


def _match_producto_evento(producto_nombre, alias_csv, titulo_evento, descripcion=''):
    """Devuelve score de match (0-100) entre un producto y un evento de calendar.

    Algoritmo:
      - Score 100: alias EXACTO encontrado en título
      - Score 80-95: alias parcial + match palabras clave
      - Score 50-79: solo palabras clave únicas matchean
      - Score < 50: no match (probable falso positivo)
    """
    if not titulo_evento:
        return 0
    texto = _normalizar(titulo_evento + ' ' + (descripcion or ''))

    # 1) Alias exact match (máxima confianza)
    if alias_csv:
        for alias in alias_csv.split(','):
            alias_norm = _normalizar(alias)
            if not alias_norm or len(alias_norm) < 3:
                continue
            # Match palabra completa con boundaries
            patron = r'(^|[^A-Z0-9+])' + _re_match.escape(alias_norm) + r'($|[^A-Z0-9+])'
            if _re_match.search(patron, texto):
                return 100

    # 2) Match por palabras clave únicas del producto
    palabras_clave = _palabras_clave_unicas(producto_nombre)
    if not palabras_clave:
        return 0

    # Cuántas palabras clave del producto aparecen en el evento
    matched = 0
    for p in palabras_clave:
        if _re_match.search(r'(^|[^A-Z0-9+])' + _re_match.escape(p) + r'($|[^A-Z0-9+])', texto):
            matched += 1
    if matched == 0:
        return 0

    # Score: porcentaje de palabras clave matcheadas
    score = int((matched / len(palabras_clave)) * 80)  # max 80 sin alias
    # Bonus si el producto tiene UNA palabra clave única y aparece
    if len(palabras_clave) == 1 and matched == 1:
        score = max(score, 75)
    return score


def _parsear_kg_evento(titulo, descripcion=''):
    """Parser AGRESIVO de kg en título/descripción.

    Detecta:
      - "90 kg", "90kg", "90KG"
      - "90,5 kg", "90.5 kg"
      - "Lote 90", "Batch 90", "90 kilos"
      - "L90", "B90"
    """
    texto = (titulo or '') + ' ' + (descripcion or '')
    # Lista de patrones de mayor a menor confianza
    patrones = [
        r'(\d+(?:[.,]\d+)?)\s*(?:kg|KG|Kg|kG)\b',           # "90 kg" - alta confianza
        r'(\d+(?:[.,]\d+)?)\s*(?:kilos?|KILOS?)\b',          # "90 kilos"
        r'\b(?:lote|LOTE|Lote|batch|BATCH|Batch)\s+(\d+(?:[.,]\d+)?)\s*(?:kg|kilos?)?', # "Lote 90"
        r'\b(?:L|B)(\d+(?:[.,]\d+)?)\s*(?:kg|kilos?)\b',     # "L90 kg"
    ]
    for patron in patrones:
        m = _re_match.search(patron, texto)
        if m:
            kg_str = m.group(1).replace(',', '.')
            try:
                v = float(kg_str)
                if 0 < v <= 5000:  # rango razonable
                    return v
            except Exception:
                pass
    return None


def _producciones_programadas_futuro(c, producto):
    """Cantidad de lotes programados a futuro (no completados)."""
    r = c.execute("""
        SELECT COALESCE(SUM(lotes), 0) FROM produccion_programada
        WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
          AND estado IN ('pendiente','en_proceso')
          AND fecha_programada >= date('now')
    """, (producto,)).fetchone()
    return int(r[0] if r else 0)


# ───────────────────────────────────────────────────────────────────────
# GENERADOR PRINCIPAL
# ───────────────────────────────────────────────────────────────────────

def generar_plan(horizonte_dias=60, tipo='auto', usuario='cron'):
    """Algoritmo central. Genera plan completo para los próximos N días.

    Devuelve dict con:
      producciones_propuestas: lista de dicts con producción a crear
      compras_propuestas: SOL automáticas de MP y envases
      conteos_propuestos: calendario conteo cíclico
      alertas: lista de alertas críticas
      log: lineas paso a paso
    """
    inicio = datetime.now()
    conn = get_db(); c = conn.cursor()

    log_lineas = []
    log_lineas.append(f"🤖 Generación auto-plan iniciada {inicio.isoformat()}")
    log_lineas.append(f"   Horizonte: {horizonte_dias}d · tipo={tipo} · usuario={usuario}")

    fecha_hoy = datetime.now().date()
    fecha_fin = fecha_hoy + timedelta(days=horizonte_dias)

    # 1. Cargar configs
    skus = c.execute("""
        SELECT spc.producto_nombre, spc.categoria, spc.cadencia_dias,
               spc.cobertura_target_dias, spc.cobertura_min_dias, spc.cobertura_max_dias,
               spc.merma_pct, spc.prioridad,
               fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
        ORDER BY spc.prioridad, spc.producto_nombre
    """).fetchall()

    log_lineas.append(f"   Productos en config: {len(skus)}")

    # 2. Para cada SKU decidir si producir y cuándo
    producciones_propuestas = []
    alertas = []
    fechas_ocupadas = {}  # area_codigo -> set de fechas ocupadas
    # Sebastian (30-abr-2026): contador de lotes proyectados POR FECHA en
    # memoria — antes el counter solo miraba BD y todas las proyecciones
    # caían el mismo día. Ahora distribuimos máx 2 lotes por día L/M/V.
    LOTES_MAX_POR_DIA = 2
    lotes_por_fecha = {}  # ISO date → count

    for sku_row in skus:
        producto, categoria, cadencia, cob_target, cob_min, cob_max, merma_pct, prioridad, lote_size_kg = sku_row
        # Sebastian (30-abr-2026): "todo debe producirse con un margen de 20
        # dias antes de que se agote" — usar 20d como umbral mínimo de
        # disparo, anulando el cob_min de la config si era mayor.
        cob_min = min(cob_min or 30, 20)
        if not lote_size_kg:
            log_lineas.append(f"   ⊘ {producto}: sin lote_size_kg en formula — saltado")
            continue

        velocidad, factor_tendencia = _velocidad_total_producto(c, producto)
        stock_actual = _stock_actual_pt(c, producto)
        ultima_prod = _ultima_produccion(c, producto)
        prog_futuro_lotes = _producciones_programadas_futuro(c, producto)

        # Velocidad ajustada por tendencia (Sebastian: "predicción agresiva")
        velocidad_proyectada = velocidad * factor_tendencia

        # Días de inventario actuales (con tendencia)
        if velocidad_proyectada > 0:
            dias_inv_actual = stock_actual / velocidad_proyectada
        else:
            dias_inv_actual = None

        # ¿Toca producir?
        razon = ''
        toca = False
        if cadencia and ultima_prod:
            dias_desde_ult = (fecha_hoy - ultima_prod).days
            if dias_desde_ult >= cadencia:
                toca = True
                razon = f'cadencia {cadencia}d cumplida ({dias_desde_ult}d desde {ultima_prod})'
        if not toca and dias_inv_actual is not None and dias_inv_actual < cob_min:
            toca = True
            razon = f'cobertura {dias_inv_actual:.0f}d < mínimo {cob_min}d'
        if not toca and dias_inv_actual is not None and dias_inv_actual < cob_target and prog_futuro_lotes == 0:
            toca = True
            razon = f'cobertura {dias_inv_actual:.0f}d < target {cob_target}d (sin prods futuras)'

        if not toca:
            continue

        # Calcular tamaño del lote: cubrir hasta cob_target con tendencia
        unidades_a_cubrir = velocidad_proyectada * cob_target
        # Aplicar merma al lote en kg
        kg_base = (lote_size_kg or 0)
        # ¿Cuántos lotes? Si la velocidad pide más de 1 lote, sugiere múltiples
        # Pero por simplicidad inicial: 1 lote = lote_size_kg, ajusta cantidad de lotes
        lotes_necesarios = 1
        # Si las unidades a cubrir × peso_g_unidad > lote_size, multiplicar lotes
        # Heurística: si velocidad × cob_target > 1.5 × velocidad × cadencia → 2 lotes
        # Por ahora: 1 lote, registramos kg con merma para que CC compre lo correcto
        kg_con_merma = kg_base * (1 + merma_pct / 100.0)

        # Asignar fecha L/M/V que no esté ocupada (BD + proyecciones en memoria)
        if ultima_prod and cadencia:
            base_fecha = ultima_prod + timedelta(days=cadencia)
        else:
            base_fecha = fecha_hoy + timedelta(days=2)
        fecha_objetivo = _next_dia_produccion(max(fecha_hoy + timedelta(days=2), base_fecha))
        # Iterar hasta encontrar día con cupo (BD + memoria), saltando L/M/V
        for _ in range(40):  # hasta 40 intentos para horizonte largo
            iso = fecha_objetivo.isoformat()
            ya_hay_bd = c.execute("""
                SELECT COUNT(*) FROM produccion_programada
                WHERE date(fecha_programada) = ? AND estado IN ('pendiente','en_proceso')
            """, (iso,)).fetchone()[0]
            ya_hay_proj = lotes_por_fecha.get(iso, 0)
            total = ya_hay_bd + ya_hay_proj
            if total < LOTES_MAX_POR_DIA:
                break
            fecha_objetivo = _next_dia_produccion(fecha_objetivo + timedelta(days=1))
        # Reservar el slot
        iso = fecha_objetivo.isoformat()
        lotes_por_fecha[iso] = lotes_por_fecha.get(iso, 0) + 1

        producciones_propuestas.append({
            'producto': producto,
            'categoria': categoria,
            'fecha_programada': fecha_objetivo.isoformat(),
            'lotes': lotes_necesarios,
            'lote_size_kg': lote_size_kg,
            'kg_con_merma': round(kg_con_merma, 2),
            'merma_pct': merma_pct,
            'velocidad_dia': round(velocidad, 2),
            'tendencia': round(factor_tendencia, 2),
            'velocidad_proyectada': round(velocidad_proyectada, 2),
            'stock_actual': stock_actual,
            'dias_inv_actual': round(dias_inv_actual, 1) if dias_inv_actual else None,
            'cobertura_target': cob_target,
            'razon': razon,
            'prioridad': prioridad,
        })
        log_lineas.append(f"   ✓ {producto[:40]:40} → {fecha_objetivo} · {kg_con_merma:.0f}kg · {razon}")

    log_lineas.append(f"   Producciones propuestas: {len(producciones_propuestas)}")

    # 3. Calcular compras anticipadas (MP + envases)
    consumo_acumulado = {}  # material_id -> g requeridos por todo el plan
    for prop in producciones_propuestas:
        items = c.execute("""
            SELECT material_id, material_nombre, COALESCE(cantidad_g_por_lote, 0), porcentaje
            FROM formula_items
            WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
        """, (prop['producto'],)).fetchall()
        for mat_id, mat_nom, cant_lote_g, pct in items:
            req_g = (cant_lote_g or 0) * prop['lotes']
            if not req_g and pct:
                req_g = (pct / 100.0) * (prop['lote_size_kg'] or 0) * 1000 * prop['lotes']
            req_g = req_g * (1 + prop['merma_pct'] / 100.0)
            consumo_acumulado[mat_id] = consumo_acumulado.get(mat_id, 0) + req_g

    compras_propuestas = []
    for mat_id, req_g in consumo_acumulado.items():
        # Stock actual
        stock = c.execute("""
            SELECT COALESCE(SUM(
                CASE WHEN tipo IN ('Ingreso','Ajuste','Devolucion') THEN cantidad
                     WHEN tipo IN ('Salida','Consumo') THEN -cantidad
                     ELSE 0 END
            ), 0) FROM movimientos WHERE material_id=?
        """, (mat_id,)).fetchone()
        stock_g = float(stock[0] if stock else 0)
        # Lead time + buffer
        lt = c.execute("""
            SELECT lead_time_dias, buffer_dias, cobertura_min_dias, cobertura_ideal_dias,
                   origen, es_envase, proveedor_principal, material_nombre
            FROM mp_lead_time_config WHERE material_id=?
        """, (mat_id,)).fetchone()
        if lt:
            lead, buffer, cob_min, cob_ideal, origen, es_envase, prov, nombre = lt
        else:
            lead, buffer, cob_min, cob_ideal, origen, es_envase, prov, nombre = 14, 30, 30, 60, 'local', 0, '', mat_id
        # ¿Falta?
        deficit_g = req_g - stock_g
        if deficit_g <= 0:
            continue
        # Cuánto pedir: deficit + cobertura ideal extra
        cantidad_a_pedir_g = deficit_g + (req_g / horizonte_dias * cob_ideal)
        # Urgencia
        if origen in ('china', 'usa', 'europa') and stock_g < req_g * 0.5:
            urgencia = 'critica'
        elif stock_g < req_g * 0.3:
            urgencia = 'critica'
        elif stock_g < req_g * 0.6:
            urgencia = 'alta'
        else:
            urgencia = 'normal'
        compras_propuestas.append({
            'material_id': mat_id,
            'material_nombre': nombre or mat_id,
            'requerido_g': round(req_g),
            'stock_actual_g': round(stock_g),
            'deficit_g': round(deficit_g),
            'cantidad_a_pedir_g': round(cantidad_a_pedir_g),
            'lead_time_dias': lead,
            'origen': origen,
            'es_envase': bool(es_envase),
            'proveedor_principal': prov,
            'urgencia': urgencia,
        })

    log_lineas.append(f"   Compras propuestas: {len(compras_propuestas)} ({sum(1 for c in compras_propuestas if c['urgencia']=='critica')} críticas)")

    # 4. Conteo cíclico para Martes/Jueves
    conteos_propuestos = []
    cursor_dia = _proximo_dia_acond(fecha_hoy + timedelta(days=1))
    materiales_a_contar = c.execute("""
        SELECT cc.material_id, cc.categoria_abc, cc.frecuencia_dias, cc.ultimo_conteo_fecha,
               COALESCE(mlt.material_nombre, cc.material_id)
        FROM conteo_ciclico_config cc
        LEFT JOIN mp_lead_time_config mlt ON mlt.material_id = cc.material_id
        ORDER BY
          CASE cc.categoria_abc WHEN 'A' THEN 0 WHEN 'B' THEN 1 ELSE 2 END,
          COALESCE(cc.ultimo_conteo_fecha, '0000-01-01') ASC
    """).fetchall()
    # Filtrar materiales que toquen contar
    contar_hoy = []
    for mat_id, cat_abc, frec, ult, nom in materiales_a_contar:
        if ult:
            try:
                ult_d = datetime.fromisoformat(ult.split('T')[0]).date()
                if (fecha_hoy - ult_d).days < (frec or 90):
                    continue
            except Exception:
                pass
        contar_hoy.append((mat_id, cat_abc, nom))
    # Distribuir uno por Ma/Ju, max 5 por día
    while contar_hoy and cursor_dia <= fecha_fin:
        if cursor_dia.weekday() in DIAS_ACOND_CONTEO:
            chunk = contar_hoy[:5]
            for mat_id, cat_abc, nom in chunk:
                conteos_propuestos.append({
                    'fecha': cursor_dia.isoformat(),
                    'material_id': mat_id,
                    'material_nombre': nom,
                    'categoria_abc': cat_abc,
                    'asignado_a': 'milton,operarios',
                })
            contar_hoy = contar_hoy[5:]
        cursor_dia = _proximo_dia_acond(cursor_dia + timedelta(days=1))

    log_lineas.append(f"   Conteos cíclicos: {len(conteos_propuestos)}")

    # 5. Compilar alertas críticas
    for cp in compras_propuestas:
        if cp['urgencia'] == 'critica':
            alertas.append({
                'tipo': 'compra_critica',
                'severidad': 'critica' if cp['origen'] == 'china' else 'alta',
                'titulo': f'⚠ {cp["material_nombre"]}: déficit {cp["deficit_g"]:.0f}g · lead {cp["lead_time_dias"]}d ({cp["origen"]})',
                'detalle': cp,
            })

    duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
    log_lineas.append(f"   ⏱ Generación completada en {duracion_ms}ms")

    return {
        'producciones_propuestas': producciones_propuestas,
        'compras_propuestas': compras_propuestas,
        'conteos_propuestos': conteos_propuestos,
        'alertas': alertas,
        'log': log_lineas,
        'duracion_ms': duracion_ms,
        'horizonte_dias': horizonte_dias,
        'fecha_hoy': fecha_hoy.isoformat(),
        'fecha_fin': fecha_fin.isoformat(),
    }


def aplicar_plan(plan, usuario='cron'):
    """Aplica el plan generado: crea produccion_programada, solicitudes_compra,
    conteo_ciclico_calendario.

    Devuelve dict con counts e ids creados.
    """
    conn = get_db(); c = conn.cursor()
    creadas_prod = []
    creadas_compras = []
    creadas_conteos = []

    for prop in plan['producciones_propuestas']:
        # Verificar si ya hay una producción para ese mismo producto+fecha
        existing = c.execute("""
            SELECT id FROM produccion_programada
            WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
              AND date(fecha_programada) = ?
              AND estado IN ('pendiente','en_proceso')
        """, (prop['producto'], prop['fecha_programada'])).fetchone()
        if existing:
            continue

        # Sumar pedidos de maquila al lote (Sebastian: "si Fernando lleva 500
        # le adiciona a la producción esas 500 unidades").
        # Considera misma fórmula (comparte_formula_con) → Kelly Guerra para Animus
        kg_extra_maquila = 0
        pedidos_maquila_asociar = []
        try:
            mq_rows = c.execute("""
                SELECT mp.id, mp.kg_estimados, mp.cliente_nombre
                FROM maquila_pedidos mp
                LEFT JOIN clientes_maquila cm ON cm.id = mp.cliente_id
                WHERE mp.estado = 'recibido'
                  AND mp.produccion_id IS NULL
                  AND (
                      UPPER(TRIM(mp.producto_nombre)) = UPPER(TRIM(?))
                   OR (cm.comparte_formula_con IS NOT NULL
                       AND UPPER(TRIM(?)) IN (
                           SELECT UPPER(TRIM(producto_nombre)) FROM formula_headers
                           WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(mp.producto_nombre))
                       ))
                  )
            """, (prop['producto'], prop['producto'])).fetchall()
            for mq in mq_rows:
                kg_extra_maquila += (mq[1] or 0)
                pedidos_maquila_asociar.append(mq[0])
        except Exception:
            pass

        kg_total = prop['kg_con_merma'] + kg_extra_maquila
        obs_extra = ''
        if kg_extra_maquila > 0:
            obs_extra = f' | Maquila: +{kg_extra_maquila:.1f}kg para {len(pedidos_maquila_asociar)} pedido(s)'

        cur = c.execute("""
            INSERT INTO produccion_programada
              (producto, fecha_programada, lotes, estado, observaciones, origen, cantidad_kg)
            VALUES (?, ?, ?, 'pendiente', ?, 'auto_plan', ?)
        """, (
            prop['producto'], prop['fecha_programada'], prop['lotes'],
            f"AUTO-PLAN ({plan['fecha_hoy']}): {prop['razon']}{obs_extra}",
            kg_total,
        ))
        creadas_prod.append(cur.lastrowid)

        # Asociar pedidos de maquila a esta producción
        for mq_id in pedidos_maquila_asociar:
            c.execute(
                "UPDATE maquila_pedidos SET produccion_id=?, estado='planificado', actualizado_en=datetime('now') WHERE id=?",
                (cur.lastrowid, mq_id)
            )

        # Asignar área automáticamente (sugerir-area)
        try:
            cap_min = (prop['kg_con_merma'] or 0) * 1.2
            tanques = c.execute("""
                SELECT codigo, area_codigo, capacidad_litros
                FROM equipos_planta
                WHERE activo=1 AND tipo IN ('tanque','marmita','olla')
                  AND capacidad_litros >= ?
                ORDER BY capacidad_litros ASC
            """, (cap_min,)).fetchall()
            por_area = {}
            for t in tanques:
                a = t[1]
                if a and (a not in por_area or t[2] < por_area[a]):
                    por_area[a] = t[2]
            for area_codigo, _cap in por_area.items():
                area_chk = c.execute(
                    "SELECT id, puede_producir, activo FROM areas_planta WHERE codigo=?",
                    (area_codigo,)
                ).fetchone()
                if area_chk and area_chk[1] and area_chk[2]:
                    c.execute("UPDATE produccion_programada SET area_id=? WHERE id=?",
                              (area_chk[0], cur.lastrowid))
                    break
        except Exception as e:
            log.warning(f'Asignación área fallida prod={cur.lastrowid}: {e}')

        # Asignar operarios automáticamente
        try:
            _asignar_operarios_a_produccion(conn, cur.lastrowid)
        except Exception as e:
            log.warning(f'Asignación operarios fallida prod={cur.lastrowid}: {e}')

    for cp in plan['compras_propuestas']:
        # Crear solicitud_compra automatizada
        # Numero unico
        next_n = c.execute("""
            SELECT COALESCE(MAX(CAST(SUBSTR(numero, 6) AS INTEGER)), 0) + 1
            FROM solicitudes_compra WHERE numero LIKE 'AUTO-%'
        """).fetchone()[0] or 1
        numero = f'AUTO-{next_n:04d}'
        cantidad_kg = round(cp['cantidad_a_pedir_g'] / 1000.0, 2)
        urgencia_solic = 'Urgente' if cp['urgencia'] == 'critica' else ('Alta' if cp['urgencia'] == 'alta' else 'Normal')
        try:
            cur = c.execute("""
                INSERT INTO solicitudes_compra
                  (numero, fecha, estado, solicitante, urgencia,
                   observaciones, area, empresa, categoria, tipo, valor)
                VALUES (?, date('now'), 'Pendiente', 'AUTO-PLAN', ?, ?, 'Producción', 'Espagiria',
                        ?, 'Compra', 0)
            """, (
                numero, urgencia_solic,
                f"AUTO-PLAN: {cp['material_nombre']} {cantidad_kg}kg · lead {cp['lead_time_dias']}d ({cp['origen']}) · déficit {cp['deficit_g']:.0f}g",
                'Materia Prima' if not cp['es_envase'] else 'Material de Empaque',
            ))
            sol_id = cur.lastrowid
            # Item asociado
            c.execute("""
                INSERT INTO solicitudes_compra_items
                  (numero, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion, valor_estimado, proveedor_sugerido)
                VALUES (?, ?, ?, ?, 'g', ?, 0, ?)
            """, (
                numero, cp['material_id'], cp['material_nombre'],
                cp['cantidad_a_pedir_g'],
                f"Para producciones plan {plan['fecha_hoy']}",
                cp.get('proveedor_principal') or '',
            ))
            creadas_compras.append({'numero': numero, 'sol_id': sol_id})
        except Exception as e:
            log.warning(f'Solicitud auto-plan fallida {cp["material_id"]}: {e}')

    for con in plan['conteos_propuestos']:
        try:
            c.execute("""
                INSERT OR IGNORE INTO conteo_ciclico_calendario
                  (fecha, material_id, material_nombre, categoria_abc, asignado_a, generado_por)
                VALUES (?, ?, ?, ?, ?, 'auto_plan')
            """, (con['fecha'], con['material_id'], con['material_nombre'],
                  con['categoria_abc'], con['asignado_a']))
            if c.rowcount:
                creadas_conteos.append(con)
        except Exception as e:
            log.warning(f'Conteo auto-plan fallido {con["material_id"]}: {e}')

    # Log de la corrida
    c.execute("""
        INSERT INTO auto_plan_runs
          (ejecutado_por, tipo, horizonte_dias, producciones_creadas,
           compras_creadas, alertas_criticas, payload_json, duracion_ms)
        VALUES (?, 'auto', ?, ?, ?, ?, ?, ?)
    """, (
        usuario, plan['horizonte_dias'],
        len(creadas_prod), len(creadas_compras), len(plan.get('alertas', [])),
        json.dumps({'log': plan['log'][:50]}, ensure_ascii=False),
        plan['duracion_ms'],
    ))
    conn.commit()

    return {
        'producciones_creadas': creadas_prod,
        'compras_creadas': creadas_compras,
        'conteos_creados': creadas_conteos,
    }


# ───────────────────────────────────────────────────────────────────────
# ENDPOINTS
# ───────────────────────────────────────────────────────────────────────

def _auth():
    if 'compras_user' not in session:
        return None, jsonify({'error': 'No autorizado'}), 401
    return session.get('compras_user', ''), None, None


@bp.route('/api/auto-plan/preview', methods=['GET'])
def auto_plan_preview():
    """Genera el plan en modo dry-run (no aplica). Sebastian ve qué propondría."""
    u, err, code = _auth()
    if err:
        return err, code
    try:
        horizonte = int(request.args.get('dias', 60))
    except Exception:
        horizonte = 60
    plan = generar_plan(horizonte_dias=horizonte, tipo='dry_run', usuario=u)
    return jsonify(plan)


@bp.route('/api/auto-plan/aplicar', methods=['POST'])
def auto_plan_aplicar():
    """Genera y aplica el plan ahora mismo (manual trigger)."""
    u, err, code = _auth()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin puede disparar el auto-plan'}), 403
    try:
        horizonte = int((request.json or {}).get('dias', 60))
    except Exception:
        horizonte = 60
    plan = generar_plan(horizonte_dias=horizonte, tipo='manual', usuario=u)
    resultado = aplicar_plan(plan, usuario=u)
    return jsonify({
        'ok': True,
        'resultado': resultado,
        'plan': {
            'producciones_propuestas': len(plan['producciones_propuestas']),
            'compras_propuestas': len(plan['compras_propuestas']),
            'conteos_propuestos': len(plan['conteos_propuestos']),
            'alertas': plan['alertas'],
            'log': plan['log'],
        },
    })


@bp.route('/api/auto-plan/ejecutar-ahora', methods=['POST'])
def auto_plan_ejecutar_ahora():
    """Dispara el cron auto-plan AHORA (sin esperar al 7am del próximo día).
    Útil para test inicial o cuando Sebastian quiere refrescar el plan."""
    u, err, code = _auth()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    try:
        from blueprints.auto_plan_jobs import ejecutar_auto_plan_diario
        from flask import current_app
        # Ejecutar en background para no bloquear el response
        import threading
        threading.Thread(
            target=ejecutar_auto_plan_diario,
            args=(current_app._get_current_object(),),
            daemon=True
        ).start()
        return jsonify({'ok': True, 'mensaje': 'Auto-plan ejecutándose en background. Mira /api/auto-plan/runs en 30s.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/api/auto-plan/asegurar-actualizado', methods=['POST'])
def auto_plan_asegurar_actualizado():
    """Sebastian (30-abr-2026): "que el cronograma sea de manera inteligente
    osea montado desde shopify, necesidades reales, monte todo con la lógica
    que hemos usado en calendario asi esta todo en la app".

    Llamada al cargar /planta. Verifica si el último auto-plan run fue hace
    más de N horas (default 12). Si sí, dispara generar+aplicar en BACKGROUND
    sin bloquear el response. El frontend muestra el plan y se refresca solo
    cuando el cron termina.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        max_horas = int(request.args.get('max_horas', 12))
    except Exception:
        max_horas = 12

    conn = get_db(); c = conn.cursor()
    last = c.execute("""
        SELECT id, ejecutado_at,
               CAST((julianday('now') - julianday(ejecutado_at)) * 24 AS INTEGER) AS horas
        FROM auto_plan_runs
        WHERE error IS NULL OR error = ''
        ORDER BY id DESC LIMIT 1
    """).fetchone()

    if last and last[2] is not None and last[2] < max_horas:
        return jsonify({
            'ok': True,
            'ejecutado': False,
            'razon': f'Plan vigente · último run hace {last[2]}h',
            'ultimo_run_at': last[1],
        })

    # Disparar en background
    try:
        from blueprints.auto_plan_jobs import ejecutar_auto_plan_diario
        from flask import current_app
        import threading
        threading.Thread(
            target=ejecutar_auto_plan_diario,
            args=(current_app._get_current_object(),),
            daemon=True
        ).start()
        return jsonify({
            'ok': True,
            'ejecutado': True,
            'mensaje': 'Auto-plan ejecutándose en background. El plan se actualizará en ~30 segundos.',
            'ultimo_run_at': last[1] if last else None,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/planta/plan-semanal-v2', methods=['GET'])
def plan_semanal_v2():
    """Sebastian (30-abr-2026): vista Semana siempre con datos.

    Si BD tiene producciones programadas → muéstralas (origen='bd').
    Si BD vacía → proyecta los próximos 14 días con el motor (origen='proyeccion').
    Cada item lleva flag de origen para que la UI muestre badge "🔮 Proyectado".

    Querystring: ?dias=14 (default)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 14))
    except Exception:
        dias = 14

    conn = get_db(); c = conn.cursor()
    fecha_desde = datetime.now().strftime('%Y-%m-%d')
    fecha_hasta = (datetime.now() + timedelta(days=dias)).strftime('%Y-%m-%d')

    # Producciones reales en BD
    rows = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes, pp.estado,
               ap.codigo as area_codigo, ap.nombre as area_nombre,
               fh.lote_size_kg, fh.imagen_url, pp.cantidad_kg
        FROM produccion_programada pp
        LEFT JOIN areas_planta ap ON ap.id = pp.area_id
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
        WHERE pp.fecha_programada BETWEEN ? AND ?
          AND pp.estado IN ('pendiente','en_proceso')
        ORDER BY pp.fecha_programada ASC, pp.id ASC
    """, (fecha_desde, fecha_hasta)).fetchall()

    items_bd = []
    for r in rows:
        items_bd.append({
            'origen': 'bd',
            'produccion_id': r[0],
            'producto': r[1],
            'fecha_programada': r[2],
            'lotes': r[3] or 1,
            'estado': r[4],
            'area_codigo': r[5],
            'area_nombre': r[6],
            'lote_size_kg': r[7],
            'imagen_url': r[8],
            'kg': r[9] or r[7] or 0,
        })

    items_proy = []
    if not items_bd:
        # BD vacía → proyectar
        proy = _generar_proyeccion_lotes(c, max(1, dias // 30 + 1))
        # Solo los próximos 'dias' días
        from datetime import datetime as _dt2
        fecha_limite = _dt2.now().date() + timedelta(days=dias)
        for p in proy:
            try:
                f = _dt2.strptime(p['fecha'], '%Y-%m-%d').date()
                if f > fecha_limite:
                    continue
            except Exception:
                continue
            # Buscar imagen del producto
            img = c.execute(
                "SELECT imagen_url FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
                (p['producto'],)
            ).fetchone()
            items_proy.append({
                'origen': 'proyeccion',
                'produccion_id': None,  # aún no existe
                'producto': p['producto'],
                'fecha_programada': p['fecha'],
                'lotes': 1,
                'estado': 'sugerido',
                'area_codigo': None,
                'area_nombre': None,
                'lote_size_kg': p['lote_kg'],
                'imagen_url': img[0] if img else None,
                'kg': p['kg_con_merma'],
                'razon': f"Cadencia {p['cadencia_dias']}d" if p.get('cadencia_dias') else 'Auto por umbral',
            })

    items = items_bd + items_proy
    items.sort(key=lambda x: x['fecha_programada'])

    # Last run info
    last_run = c.execute("""
        SELECT ejecutado_at,
               CAST((julianday('now') - julianday(ejecutado_at)) * 24 AS INTEGER) AS horas
        FROM auto_plan_runs ORDER BY id DESC LIMIT 1
    """).fetchone()

    return jsonify({
        'rango': {'desde': fecha_desde, 'hasta': fecha_hasta, 'dias': dias},
        'items': items,
        'kpis': {
            'total': len(items),
            'desde_bd': len(items_bd),
            'proyectadas': len(items_proy),
        },
        'auto_plan_status': {
            'ultimo_run_at': last_run[0] if last_run else None,
            'horas_desde_run': last_run[1] if last_run and last_run[1] is not None else None,
            'plan_vacio': len(items_bd) == 0,
        },
    })


@bp.route('/api/planta/confirmar-proyeccion', methods=['POST'])
def confirmar_proyeccion():
    """Cuando el operario ve una proyección y la quiere persistir como
    producción real, este endpoint la mete en produccion_programada.
    Body: {producto, fecha_programada, lotes?, lote_size_kg?, kg?}"""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    producto = (d.get('producto') or '').strip()
    fecha = (d.get('fecha_programada') or '').strip()
    if not producto or not fecha:
        return jsonify({'error': 'producto y fecha_programada requeridos'}), 400
    conn = get_db(); c = conn.cursor()
    # Verificar duplicado
    existing = c.execute("""
        SELECT id FROM produccion_programada
        WHERE UPPER(TRIM(producto))=UPPER(TRIM(?))
          AND date(fecha_programada)=date(?)
          AND estado IN ('pendiente','en_proceso')
    """, (producto, fecha)).fetchone()
    if existing:
        return jsonify({'ok': True, 'id': existing[0], 'ya_existia': True})
    kg_total = float(d.get('kg') or d.get('lote_size_kg') or 0)
    cur = c.execute("""
        INSERT INTO produccion_programada
          (producto, fecha_programada, lotes, estado, observaciones, origen, cantidad_kg)
        VALUES (?, ?, ?, 'pendiente', ?, 'confirmacion_manual', ?)
    """, (
        producto, fecha, int(d.get('lotes') or 1),
        f'Confirmado desde proyección por {user} el {datetime.now().isoformat()}',
        kg_total,
    ))
    nuevo_id = cur.lastrowid
    # Auto-asignar área + operarios
    try:
        if kg_total > 0:
            cap_min = kg_total * 1.2
            tanques = c.execute("""
                SELECT area_codigo FROM equipos_planta
                WHERE activo=1 AND tipo IN ('tanque','marmita','olla')
                  AND capacidad_litros >= ?
                ORDER BY capacidad_litros ASC LIMIT 1
            """, (cap_min,)).fetchone()
            if tanques:
                area_chk = c.execute(
                    "SELECT id FROM areas_planta WHERE codigo=? AND puede_producir=1 AND activo=1",
                    (tanques[0],)
                ).fetchone()
                if area_chk:
                    c.execute("UPDATE produccion_programada SET area_id=? WHERE id=?",
                              (area_chk[0], nuevo_id))
        _asignar_operarios_a_produccion(conn, nuevo_id)
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'id': nuevo_id, 'ya_existia': False})


@bp.route('/api/planta/produccion/<int:prod_id>/eliminar-y-replanificar', methods=['POST'])
def eliminar_y_replan(prod_id):
    """Sebastian (30-abr-2026): "si ya la hicimos le demos eliminar, y ponga
    otra alli automaticamente".

    Elimina la producción y propone otra del MISMO producto en la próxima
    fecha disponible según cadencia/cobertura.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    motivo = (d.get('motivo') or 'Eliminada manualmente').strip()
    conn = get_db(); c = conn.cursor()

    # Obtener datos de la producción antes de borrarla
    pp = c.execute("""
        SELECT producto, fecha_programada, lotes, cantidad_kg
        FROM produccion_programada WHERE id=?
    """, (prod_id,)).fetchone()
    if not pp:
        return jsonify({'error': 'Producción no existe'}), 404
    producto, fecha_orig, lotes, kg = pp

    # Marcar como cancelada (no eliminamos para no perder histórico)
    c.execute("""
        UPDATE produccion_programada SET
          estado='cancelado',
          observaciones=COALESCE(observaciones,'')||' | Cancelada por '||?||': '||?
        WHERE id=?
    """, (user, motivo, prod_id))

    # Calcular próxima fecha sugerida según cadencia
    cfg = c.execute("""
        SELECT cadencia_dias, cobertura_target_dias, cobertura_min_dias, merma_pct
        FROM sku_planeacion_config
        WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))
    """, (producto,)).fetchone()
    cadencia = cfg[0] if cfg else None
    cob_target = cfg[1] if cfg else 60

    velocidad, factor = _velocidad_total_producto(c, producto)
    velocidad_proj = max(0.5, velocidad * factor)
    stock_actual = _stock_actual_pt(c, producto)

    # Si tiene cadencia → próxima = fecha_orig + cadencia
    # Si no → cuando bajará bajo umbral 20d
    from datetime import datetime as _dt2
    fecha_orig_dt = _dt2.strptime(fecha_orig[:10], '%Y-%m-%d').date() if fecha_orig else _dt2.now().date()
    if cadencia:
        proxima = fecha_orig_dt + timedelta(days=cadencia)
    else:
        # Días hasta bajar de 20
        dias_disp = max(0, (stock_actual / velocidad_proj) - 20)
        proxima = _dt2.now().date() + timedelta(days=int(dias_disp))
    # Asegurar L/M/V
    while proxima.weekday() not in (0, 2, 4):
        proxima += timedelta(days=1)

    # Crear nueva producción sugerida
    cur = c.execute("""
        INSERT INTO produccion_programada
          (producto, fecha_programada, lotes, estado, observaciones, origen, cantidad_kg)
        VALUES (?, ?, ?, 'pendiente', ?, 'replan_post_eliminacion', ?)
    """, (
        producto, proxima.isoformat(), lotes or 1,
        f'Replanificada tras cancelación de id={prod_id}',
        kg or 0,
    ))
    nuevo_id = cur.lastrowid
    conn.commit()
    return jsonify({
        'ok': True,
        'eliminado': prod_id,
        'nueva_id': nuevo_id,
        'nueva_fecha': proxima.isoformat(),
        'producto': producto,
        'mensaje': f'Producción cancelada · nueva sugerida para {proxima.isoformat()}',
    })


@bp.route('/api/planta/produccion/<int:prod_id>/editar-lote', methods=['POST'])
def editar_lote(prod_id):
    """Sebastian (30-abr-2026): "permita editar la cantidad del lote y
    calcule todo".

    Body: {cantidad_kg: float, lotes?: int}
    Recalcula MP requerida + envases + costos al cambiar tamaño.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    nueva_kg = d.get('cantidad_kg')
    if nueva_kg is None:
        return jsonify({'error': 'cantidad_kg requerido'}), 400
    nueva_kg = float(nueva_kg)
    nuevos_lotes = int(d.get('lotes') or 1)
    conn = get_db(); c = conn.cursor()

    pp = c.execute(
        "SELECT producto, lotes, cantidad_kg FROM produccion_programada WHERE id=?",
        (prod_id,)
    ).fetchone()
    if not pp:
        return jsonify({'error': 'Producción no existe'}), 404
    producto, lotes_actuales, kg_actuales = pp

    c.execute("""
        UPDATE produccion_programada SET
          cantidad_kg=?, lotes=?,
          observaciones=COALESCE(observaciones,'')||' | Lote editado por '||?||' '||datetime('now')||': '||?||'kg→'||?||'kg'
        WHERE id=?
    """, (nueva_kg, nuevos_lotes, user, kg_actuales or 0, nueva_kg, prod_id))
    conn.commit()

    # Recalcular MP requerida + envases con el nuevo tamaño
    fh = c.execute(
        "SELECT lote_size_kg FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    lote_kg_base = (fh[0] if fh else 0) or 0
    factor = nueva_kg / lote_kg_base if lote_kg_base else 1
    items = c.execute("""
        SELECT material_id, material_nombre, COALESCE(cantidad_g_por_lote,0)
        FROM formula_items
        WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))
    """, (producto,)).fetchall()
    mp_recalc = []
    for mat_id, mat_nom, cant_g in items:
        nueva_g = (cant_g or 0) * factor * nuevos_lotes
        mp_recalc.append({'material_id': mat_id, 'material_nombre': mat_nom,
                          'gramos_requeridos': round(nueva_g)})

    # Envases recalculados
    pres = c.execute("""
        SELECT envase_codigo, factor_g_por_unidad, etiqueta
        FROM producto_presentaciones
        WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND activo=1 AND es_default=1
        LIMIT 1
    """, (producto,)).fetchone()
    envase_recalc = None
    if pres and pres[1] and pres[1] > 0:
        unidades = (nueva_kg * 1000 * nuevos_lotes) / pres[1]
        envase_recalc = {
            'envase_codigo': pres[0],
            'etiqueta': pres[2],
            'unidades_requeridas': int(unidades),
        }

    return jsonify({
        'ok': True,
        'produccion_id': prod_id,
        'producto': producto,
        'cantidad_kg_anterior': kg_actuales,
        'cantidad_kg_nueva': nueva_kg,
        'mp_recalculada': mp_recalc,
        'envase_recalculado': envase_recalc,
    })


@bp.route('/api/planta/detectar-cambios-demanda', methods=['GET'])
def detectar_cambios_demanda():
    """Sebastian (30-abr-2026): "si algo debe modificarse que me diga de
    manera inmediata aumento venta aparezca alli recomiendo mover a tal
    día y este a otro deseas aceptar?".

    Compara velocidad actual (últimos 14d) vs base (días 15-44 atrás).
    Detecta SKUs con cambio significativo (>20% arriba o abajo).
    Sugiere ajustes al calendario.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    skus_config = c.execute("""
        SELECT producto_nombre FROM sku_planeacion_config WHERE activo=1
    """).fetchall()

    cambios_detectados = []
    for (producto,) in skus_config:
        # Velocidad reciente (14d) vs base (15-44d atrás)
        sku_rows = c.execute(
            "SELECT sku FROM sku_producto_map WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND COALESCE(activo,1)=1",
            (producto,)
        ).fetchall()
        if not sku_rows:
            continue
        v_reciente_total = 0.0
        v_base_total = 0.0
        for (sku,) in sku_rows:
            # Reciente
            r = _ventas_diarias_por_sku(c, sku, dias=14)
            if r:
                v_reciente_total += sum(q for _, q in r) / 14.0
            # Base (días 15-44)
            r2 = c.execute("""
                SELECT date(creado_en), sku_items
                FROM animus_shopify_orders
                WHERE creado_en BETWEEN date('now','-44 days') AND date('now','-15 days')
                  AND sku_items IS NOT NULL
            """).fetchall()
            cant = 0
            for fecha, sku_items_json in r2:
                if not sku_items_json:
                    continue
                try:
                    items = json.loads(sku_items_json) if isinstance(sku_items_json, str) else sku_items_json
                except Exception:
                    continue
                for it in (items or []):
                    if (it.get('sku') or '').strip() == sku:
                        cant += float(it.get('cantidad') or it.get('quantity') or 0)
            if cant > 0:
                v_base_total += cant / 30.0
        if v_base_total < 0.1 or v_reciente_total < 0.1:
            continue
        cambio_pct = ((v_reciente_total - v_base_total) / v_base_total) * 100
        if abs(cambio_pct) < 20:
            continue
        # Cambio significativo
        # Buscar próxima producción programada
        prox = c.execute("""
            SELECT id, fecha_programada FROM produccion_programada
            WHERE UPPER(TRIM(producto))=UPPER(TRIM(?))
              AND fecha_programada >= date('now')
              AND estado IN ('pendiente','en_proceso')
            ORDER BY fecha_programada ASC LIMIT 1
        """, (producto,)).fetchone()
        recomendacion = ''
        nueva_fecha = None
        if cambio_pct > 0 and prox:
            # Aumento: adelantar
            try:
                fp = datetime.strptime(prox[1][:10], '%Y-%m-%d').date()
                # Adelantar tantos días como porcentaje × 0.3 (heurística)
                dias_adelanto = min(14, int(cambio_pct * 0.2))
                nueva_fecha = fp - timedelta(days=dias_adelanto)
                while nueva_fecha.weekday() not in (0, 2, 4):
                    nueva_fecha -= timedelta(days=1)
                if nueva_fecha < datetime.now().date():
                    nueva_fecha = datetime.now().date()
                    while nueva_fecha.weekday() not in (0, 2, 4):
                        nueva_fecha += timedelta(days=1)
                recomendacion = f'Adelantar de {prox[1]} a {nueva_fecha.isoformat()}'
            except Exception:
                pass
        cambios_detectados.append({
            'producto': producto,
            'velocidad_base': round(v_base_total, 2),
            'velocidad_reciente': round(v_reciente_total, 2),
            'cambio_pct': round(cambio_pct, 1),
            'tipo': 'aumento' if cambio_pct > 0 else 'caida',
            'severidad': 'alta' if abs(cambio_pct) > 50 else 'media',
            'proxima_produccion_id': prox[0] if prox else None,
            'proxima_fecha_actual': prox[1] if prox else None,
            'fecha_sugerida': nueva_fecha.isoformat() if nueva_fecha else None,
            'recomendacion': recomendacion,
        })

    return jsonify({
        'cambios': cambios_detectados,
        'total': len(cambios_detectados),
    })


# ════════════════════════════════════════════════════════════════════════
# DEBUG CALENDAR — visibilidad total de qué se lee y cómo se matchea
# ════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════
# RECOMENDACIONES INTELIGENTES — la lógica clara
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "tenemos 30 SKUs, una venta diaria y un stock que
# puedes ver en shopify, en calendar aparece cada cuanto los hago normalmente,
# cuántos kilos, cuántos hice ya entre la semana pasada y esta que aún no
# entran al inventario de Shopify · con esa lógica tú podrías decirme el
# próximo lunes deberia producirse tal producto".

def _proximo_dia_lmv(desde_fecha):
    """Próximo L/M/V desde una fecha base."""
    f = desde_fecha
    while f.weekday() not in (0, 2, 4):
        f += timedelta(days=1)
    return f


def _clasificar_estado_sku(c, producto, velocidad_dia, ultima_prod_calendar):
    """Clasifica el SKU automáticamente:
       - sin_ventas: 0 ventas en 90 días Y sin produccion calendar reciente
       - baja_rotacion: velocidad < 0.05 u/día (prácticamente no se vende)
       - activo: caso normal
    Sebastian: "estos en rojo ya no los producimos varias cosas".
    """
    # Si la velocidad es prácticamente 0 (< 0.05/día = menos de 1.5 unid/mes)
    # Y no se produjo en últimos 60 días en calendar → sin_ventas
    fecha_hoy = datetime.now().date()
    if velocidad_dia < 0.05:
        if ultima_prod_calendar:
            dias_desde = (fecha_hoy - ultima_prod_calendar).days
            if dias_desde > 90:
                return 'sin_ventas', f'Velocidad {velocidad_dia:.3f} u/día y no producido en {dias_desde}d'
            else:
                return 'baja_rotacion', f'Velocidad muy baja {velocidad_dia:.3f} u/día (rotación lenta)'
        else:
            return 'sin_ventas', f'Velocidad {velocidad_dia:.3f} u/día y nunca producido en Calendar'
    return 'activo', None


def _calcular_recomendacion_sku(c, producto, lote_kg_default, cadencia_cfg, merma_pct, prioridad, ignorar_calendar=False):
    """LA función central. Para un SKU devuelve la recomendación completa.

    Lógica:
      1. Stock total = Shopify stock + Pipeline (calendar últimos 14d)
                                     + Futuro programado (calendar/BD futuro)
      2. Velocidad = ventas Shopify últimos 30d / 30
      3. Días alcance = stock_total / velocidad
      4. Si alcance > target+margen → OK no producir
      5. Si alcance > margen → producir cuando falten margen días
      6. Si alcance ≤ margen → URGENTE
      7. Cadencia histórica de calendar > 14d atrás como referencia
      8. Lote típico = mediana de kg en calendar histórico

    Si ignorar_calendar=True: NO suma pipeline_kg ni futuro_kg de Calendar
    al stock_total (solo BD futuro). Sebastian (30-abr-2026): "no tengas en
    cuenta el calendario, si tuvieras que producir ya según Shopify".
    """
    fecha_hoy = datetime.now().date()
    fecha_pipeline_inicio = fecha_hoy - timedelta(days=14)
    margen_dias = 20  # Sebastian: producir 20d antes de agotar
    cobertura_target = 60  # objetivo

    # 1. Velocidad de venta (Shopify 30d con tendencia)
    velocidad, factor = _velocidad_total_producto(c, producto)
    velocidad_proj = max(0.01, velocidad * factor)

    # 2. Stock actual (Shopify)
    stock_shopify = _stock_actual_pt(c, producto)

    # 3. Pipeline: producciones del calendar entre hoy-14d y hoy (ya hechas
    #    pero aún no entran a Shopify)
    eventos = _calendar_events_cached()
    alias = _alias_calendar_for(c, producto)
    factor_g = _factor_g_por_unidad(c, producto)
    pipeline_kg = 0.0
    futuro_kg = 0.0
    historico_fechas = []
    historico_kg_list = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
        if score < 60:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', '')) or lote_kg_default or 30
        if fecha_pipeline_inicio <= f <= fecha_hoy:
            # Pipeline (no entra aún a Shopify pero ya producido)
            pipeline_kg += kg
        elif f > fecha_hoy:
            # Futuro programado
            futuro_kg += kg
        else:
            # Histórico (>14d atrás)
            historico_fechas.append(f)
            historico_kg_list.append(kg)

    # También leer producciones futuras de BD
    bd_futuro = c.execute("""
        SELECT cantidad_kg FROM produccion_programada
        WHERE UPPER(TRIM(producto))=UPPER(TRIM(?))
          AND estado IN ('pendiente','en_proceso')
          AND fecha_programada >= date('now')
    """, (producto,)).fetchall()
    bd_futuro_kg = sum(float(r[0] or 0) for r in bd_futuro)
    futuro_kg += bd_futuro_kg

    pipeline_unidades = (pipeline_kg * 1000) / max(factor_g, 1)
    futuro_unidades = (futuro_kg * 1000) / max(factor_g, 1)
    if ignorar_calendar:
        # Solo BD futuro cuenta (lo que Sebastián confirmó manualmente),
        # ignoramos pipeline (Calendar últimos 14d) y futuro de Calendar.
        bd_futuro_unidades = (bd_futuro_kg * 1000) / max(factor_g, 1)
        stock_total = stock_shopify + bd_futuro_unidades
        # Para reportar al frontend con valores limpios:
        pipeline_kg = 0.0
        pipeline_unidades = 0
        futuro_kg = bd_futuro_kg
        futuro_unidades = int(bd_futuro_unidades)
    else:
        stock_total = stock_shopify + pipeline_unidades + futuro_unidades
    dias_alcance = stock_total / velocidad_proj if velocidad_proj > 0 else 999

    # 4. Cadencia y lote típico desde histórico calendar
    cadencia_real = _calcular_cadencia_real(historico_fechas) if len(historico_fechas) >= 2 else None
    lote_tipico_kg = sorted(historico_kg_list)[len(historico_kg_list)//2] if historico_kg_list else (lote_kg_default or 30)

    # 4.5. Estado del SKU (descontinuado / sin_ventas / activo)
    estado_db = c.execute(
        "SELECT estado, razon_estado FROM sku_planeacion_config WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    estado_actual = (estado_db[0] if estado_db else 'activo') or 'activo'
    ultima_prod_cal = max(historico_fechas) if historico_fechas else None
    # Si está marcado como descontinuado/pausado por el usuario, respetarlo
    if estado_actual in ('descontinuado', 'pausado'):
        urgencia = 'inactivo'
        fecha_proxima = None
        razon = f'{estado_actual.upper()} · ' + (estado_db[1] if estado_db and estado_db[1] else 'Marcado por usuario')
    else:
        # Auto-clasificar
        nuevo_estado, razon_auto = _clasificar_estado_sku(c, producto, velocidad_proj, ultima_prod_cal)
        if nuevo_estado == 'sin_ventas':
            urgencia = 'sin_ventas'
            fecha_proxima = None
            razon = razon_auto
        elif nuevo_estado == 'baja_rotacion' and stock_total > 50:
            urgencia = 'baja_rotacion'
            fecha_proxima = None
            razon = f'{razon_auto}. Stock {stock_total:.0f}u suficiente.'
        # 5. Recomendación normal
        elif dias_alcance > cobertura_target + margen_dias:
            urgencia = 'innecesaria'
            fecha_proxima = None
            razon = f'Stock cubre {dias_alcance:.0f}d (objetivo {cobertura_target}d). No urgente.'
        elif dias_alcance > margen_dias:
            dias_hasta = max(2, int(dias_alcance - margen_dias))
            fecha_proxima = _proximo_dia_lmv(fecha_hoy + timedelta(days=dias_hasta))
            urgencia = 'baja' if dias_alcance > 30 else 'media'
            razon = f'Stock alcanza {dias_alcance:.0f}d. Producir antes de bajar a margen 20d.'
        elif dias_alcance > 7:
            fecha_proxima = _proximo_dia_lmv(fecha_hoy + timedelta(days=2))
            urgencia = 'alta'
            razon = f'Stock alcanza solo {dias_alcance:.0f}d (margen mínimo 20d). Producir esta semana.'
        else:
            fecha_proxima = _proximo_dia_lmv(fecha_hoy + timedelta(days=1))
            urgencia = 'critica'
            razon = f'STOCK CRÍTICO · solo {dias_alcance:.0f}d. Producir ya.'

    # Considerar cadencia: si hace mucho que no se produce y aún hay stock, igual
    # programar pronto (consistencia)
    if cadencia_real and historico_fechas:
        ultima = max(historico_fechas)
        dias_desde_ultima = (fecha_hoy - ultima).days
        if dias_desde_ultima > cadencia_real * 1.2 and urgencia == 'innecesaria':
            urgencia = 'baja'
            fecha_proxima = _proximo_dia_lmv(fecha_hoy + timedelta(days=7))
            razon = f'{dias_desde_ultima}d desde última producción (cadencia {cadencia_real}d). Ya toca por consistencia.'

    return {
        'producto': producto,
        'velocidad_dia': round(velocidad_proj, 2),
        'stock_shopify': int(stock_shopify),
        'pipeline_kg': round(pipeline_kg, 1),
        'pipeline_unidades': int(pipeline_unidades),
        'futuro_kg': round(futuro_kg, 1),
        'futuro_unidades': int(futuro_unidades),
        'stock_total_unidades': int(stock_total),
        'dias_alcance': round(dias_alcance, 1),
        'cadencia_historica_dias': cadencia_real,
        'cadencia_configurada': cadencia_cfg,
        'lote_tipico_kg': round(lote_tipico_kg, 1),
        'lote_kg_default': lote_kg_default,
        'producciones_historicas': len(historico_fechas),
        'ultima_produccion': max(historico_fechas).isoformat() if historico_fechas else None,
        'fecha_proxima': fecha_proxima.isoformat() if fecha_proxima else None,
        'fecha_proxima_dia_semana': fecha_proxima.strftime('%A') if fecha_proxima else None,
        'urgencia': urgencia,
        'razon': razon,
        'factor_g': factor_g,
        'estado_sku': estado_actual,
    }


@bp.route('/api/planta/sku/<int:sku_id>/estado', methods=['POST'])
def actualizar_estado_sku(sku_id):
    """Marca un SKU como descontinuado/pausado/activo."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    nuevo = (d.get('estado') or '').strip()
    if nuevo not in ('activo', 'descontinuado', 'pausado'):
        return jsonify({'error': 'estado inválido'}), 400
    razon = (d.get('razon') or '').strip() or None
    conn = get_db(); c = conn.cursor()
    if nuevo == 'activo':
        c.execute("""
            UPDATE sku_planeacion_config SET estado='activo',
              descontinuado_at=NULL, descontinuado_por=NULL, razon_estado=NULL,
              actualizado_en=datetime('now')
            WHERE id=?
        """, (sku_id,))
    else:
        c.execute("""
            UPDATE sku_planeacion_config SET estado=?,
              descontinuado_at=datetime('now'), descontinuado_por=?, razon_estado=?,
              actualizado_en=datetime('now')
            WHERE id=?
        """, (nuevo, user, razon, sku_id))
    conn.commit()
    return jsonify({'ok': True, 'estado': nuevo})


@bp.route('/api/planta/recomendaciones', methods=['GET'])
def planta_recomendaciones():
    """Devuelve recomendaciones inteligentes por SKU.
    Sebastian: "tú podrías decirme el próximo lunes deberia producirse tal producto".

    Query param ?ignorar_calendar=1 → modo "puro Shopify": ignora pipeline+futuro
    de Calendar al sumar stock_total. Útil cuando el Calendar no es confiable.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    ignorar_cal = request.args.get('ignorar_calendar', '0') in ('1', 'true', 'yes')
    conn = get_db(); c = conn.cursor()
    skus = c.execute("""
        SELECT spc.producto_nombre, spc.cadencia_dias, spc.merma_pct, spc.prioridad,
               fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
        ORDER BY spc.prioridad ASC
    """).fetchall()
    recomendaciones = []
    for sku_row in skus:
        producto, cadencia, merma, prioridad, lote_kg = sku_row
        try:
            rec = _calcular_recomendacion_sku(c, producto, lote_kg, cadencia, merma, prioridad, ignorar_calendar=ignorar_cal)
            recomendaciones.append(rec)
        except Exception as e:
            log.warning(f'Recomendación fallida {producto}: {e}')

    # Sebastian (30-abr-2026): "ideal una producción por día, si hay semanas
    # donde puede ser L-M-V mejor, depende de cómo de" — distribución
    # inteligente: si hay >3 lotes urgentes/semana → distribuir L-V (5 días);
    # si <=3 → L/M/V (3 días).
    accionables_count = sum(1 for r in recomendaciones if r['urgencia'] in ('critica', 'alta', 'media'))
    distribuir_lv_completo = accionables_count > 3
    DIAS_PRODUCCION = (0, 1, 2, 3, 4) if distribuir_lv_completo else (0, 2, 4)

    # Re-distribuir fechas de las accionables para no saturar
    fechas_ocupadas_count = {}
    LIMITE_POR_DIA = 1
    fecha_hoy = datetime.now().date()

    def _proximo_dia_disponible(desde, dias_validos):
        f = max(desde, fecha_hoy + timedelta(days=2))
        for _ in range(60):
            while f.weekday() not in dias_validos:
                f += timedelta(days=1)
            iso = f.isoformat()
            if fechas_ocupadas_count.get(iso, 0) < LIMITE_POR_DIA:
                return f
            f += timedelta(days=1)
        return f

    # Ordenar por urgencia
    orden_urg = {'critica': 0, 'alta': 1, 'media': 2, 'baja': 3, 'innecesaria': 4,
                 'sin_ventas': 5, 'baja_rotacion': 5, 'inactivo': 5}
    recomendaciones.sort(key=lambda r: (orden_urg.get(r['urgencia'], 6), -float(r.get('dias_alcance') or 9999)))

    # Re-asignar fechas próximas para los accionables (críticos primero)
    for r in recomendaciones:
        if r['urgencia'] not in ('critica', 'alta', 'media', 'baja') or not r.get('fecha_proxima'):
            continue
        try:
            fecha_orig = datetime.strptime(r['fecha_proxima'], '%Y-%m-%d').date()
        except Exception:
            fecha_orig = fecha_hoy + timedelta(days=2)
        nueva = _proximo_dia_disponible(fecha_orig, DIAS_PRODUCCION)
        r['fecha_proxima'] = nueva.isoformat()
        r['fecha_proxima_dia_semana'] = nueva.strftime('%A')
        iso = nueva.isoformat()
        fechas_ocupadas_count[iso] = fechas_ocupadas_count.get(iso, 0) + 1

    # KPIs
    kpis = {
        'total': len(recomendaciones),
        'criticas': sum(1 for r in recomendaciones if r['urgencia'] == 'critica'),
        'altas': sum(1 for r in recomendaciones if r['urgencia'] == 'alta'),
        'medias': sum(1 for r in recomendaciones if r['urgencia'] == 'media'),
        'bajas': sum(1 for r in recomendaciones if r['urgencia'] == 'baja'),
        'innecesarias': sum(1 for r in recomendaciones if r['urgencia'] == 'innecesaria'),
        'sin_ventas': sum(1 for r in recomendaciones if r['urgencia'] == 'sin_ventas'),
        'baja_rotacion': sum(1 for r in recomendaciones if r['urgencia'] == 'baja_rotacion'),
        'inactivos': sum(1 for r in recomendaciones if r['urgencia'] == 'inactivo'),
    }
    return jsonify({
        'recomendaciones': recomendaciones,
        'kpis': kpis,
        'fecha_analisis': datetime.now().isoformat(),
        'modo': 'shopify_puro' if ignorar_cal else 'completo',
        'ignorar_calendar': ignorar_cal,
        'distribucion': {
            'patron': 'L-V (5 días)' if distribuir_lv_completo else 'L/M/V (3 días)',
            'razon': f'{accionables_count} producciones accionables → ' + ('alta carga, distribución 5 días' if distribuir_lv_completo else 'carga normal, L/M/V suficiente'),
            'limite_por_dia': LIMITE_POR_DIA,
        },
    })


# ════════════════════════════════════════════════════════════════════════
# DIAGNÓSTICO SKU — qué LEE el sistema crudo de Shopify para un producto
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "no lo veo en la app no veo que lo esté
# calculando". Visibilidad total: cada número que el motor usa, expuesto.

@bp.route('/api/planta/diagnostico-sku', methods=['GET'])
def diagnostico_sku():
    """Devuelve TODO lo que el sistema lee crudo para un producto:

      - SKUs Shopify mapeados (SAH, SAH10...)
      - Stock por SKU + total
      - Pedidos últimos 30/60/90/365 días
      - Velocidad calculada (con tendencia)
      - Factor g/u que usa el motor
      - Días de alcance hoy
      - Lote típico histórico
      - Datos crudos de las últimas N transacciones para verificar

    Query params:
      - producto: nombre exacto del producto
      - listar: si '1' devuelve solo lista de productos disponibles
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    conn = get_db(); c = conn.cursor()

    # Modo lista: solo nombres de productos
    if request.args.get('listar') == '1':
        rows = c.execute("""
            SELECT producto_nombre, COALESCE(estado,'activo'), prioridad
            FROM sku_planeacion_config
            WHERE activo = 1
            ORDER BY prioridad ASC, producto_nombre ASC
        """).fetchall()
        return jsonify({
            'productos': [{'nombre': r[0], 'estado': r[1], 'prioridad': r[2]} for r in rows]
        })

    producto = (request.args.get('producto') or '').strip()
    if not producto:
        return jsonify({'error': 'Falta param "producto"'}), 400

    out = {
        'producto': producto,
        'fecha_analisis': datetime.now().isoformat(),
        'timestamp_actual': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    # 1) SKUs Shopify mapeados a este producto
    skus_map = c.execute("""
        SELECT sku, COALESCE(activo,1) FROM sku_producto_map
        WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
        ORDER BY sku
    """, (producto,)).fetchall()
    out['skus_mapeados'] = [{'sku': r[0], 'activo': bool(r[1])} for r in skus_map]
    out['advertencias'] = []
    if not skus_map:
        out['advertencias'].append(f'⚠ No hay SKUs Shopify mapeados a este producto en sku_producto_map. Sin mapeo, NO se pueden leer ventas ni stock.')

    # 2) Stock por SKU desde stock_pt
    stock_detalle = []
    stock_total = 0
    for (sku, activo) in skus_map:
        try:
            rs = c.execute("""
                SELECT lote_produccion, fecha_produccion, unidades_inicial,
                       unidades_disponible, COALESCE(estado,'OK')
                FROM stock_pt
                WHERE sku = ?
                ORDER BY fecha_produccion DESC
            """, (sku,)).fetchall()
        except Exception as e:
            stock_detalle.append({'sku': sku, 'error': str(e), 'lotes': []})
            continue
        lotes_sku = []
        total_sku = 0
        for r in rs:
            disp = int(r[3] or 0)
            lotes_sku.append({
                'lote': r[0],
                'fecha': r[1],
                'inicial': int(r[2] or 0),
                'disponible': disp,
                'estado': r[4],
            })
            if (r[4] or 'OK').upper() != 'AGOTADO':
                total_sku += disp
        stock_detalle.append({
            'sku': sku,
            'total_unidades': total_sku,
            'lotes_count': len(lotes_sku),
            'lotes': lotes_sku[:10],
        })
        stock_total += total_sku
    out['stock_por_sku'] = stock_detalle
    out['stock_total_unidades'] = stock_total

    if stock_total == 0 and skus_map:
        out['advertencias'].append('⚠ Stock = 0. Verifica que stock_pt tenga datos sincronizados con Shopify.')

    # 3) Ventas últimos N días por SKU
    ventas_por_periodo = {}
    for dias in [30, 60, 90, 365]:
        total_periodo = 0
        detalle_sku = {}
        for (sku, _) in skus_map:
            ventas = _ventas_diarias_por_sku(c, sku, dias=dias)
            unidades_sku = sum(q for _, q in ventas)
            detalle_sku[sku] = {
                'unidades': int(unidades_sku),
                'dias_con_venta': len(ventas),
                'velocidad_dia': round(unidades_sku / dias, 3) if dias > 0 else 0,
            }
            total_periodo += unidades_sku
        ventas_por_periodo[f'{dias}d'] = {
            'total_unidades': int(total_periodo),
            'velocidad_promedio': round(total_periodo / dias, 3) if dias > 0 else 0,
            'por_sku': detalle_sku,
        }
    out['ventas_por_periodo'] = ventas_por_periodo

    # 4) Velocidad final con tendencia (lo que el motor usa)
    velocidad, factor = _velocidad_total_producto(c, producto)
    velocidad_proj = max(0.0, velocidad * factor)
    out['velocidad_final'] = {
        'velocidad_base': round(velocidad, 3),
        'factor_tendencia': round(factor, 3),
        'velocidad_ajustada': round(velocidad_proj, 3),
        'unidades_por_dia': round(velocidad_proj, 2),
        'unidades_por_semana': round(velocidad_proj * 7, 1),
        'unidades_por_mes': round(velocidad_proj * 30, 1),
    }

    # 5) Factor g/u
    factor_g = _factor_g_por_unidad(c, producto)
    out['factor_g_por_unidad'] = factor_g

    # 6) Días de alcance HOY (sin Calendar)
    if velocidad_proj > 0.01:
        dias_alcance = stock_total / velocidad_proj
        out['dias_alcance_hoy'] = round(dias_alcance, 1)
        out['fecha_stockout_proyectada'] = (datetime.now().date() + timedelta(days=int(dias_alcance))).isoformat()
        margen = 20
        if dias_alcance > margen:
            dias_hasta_lote = int(dias_alcance - margen)
            out['fecha_lote_recomendada'] = (datetime.now().date() + timedelta(days=dias_hasta_lote)).isoformat()
            out['urgencia'] = 'OK · falta(n) ' + str(dias_hasta_lote) + 'd para programar'
        else:
            out['fecha_lote_recomendada'] = (datetime.now().date() + timedelta(days=2)).isoformat()
            out['urgencia'] = 'URGENTE · stock alcanza ' + str(round(dias_alcance, 1)) + 'd, debió producirse hace ' + str(margen - int(dias_alcance)) + 'd'
    else:
        out['dias_alcance_hoy'] = None
        out['urgencia'] = 'SIN VENTAS · velocidad ~0, no se planea'

    # 7) Lote típico (mediana histórica Calendar > 14d)
    eventos = _calendar_events_cached()
    alias = _alias_calendar_for(c, producto)
    fecha_pip = datetime.now().date() - timedelta(days=14)
    kgs_hist = []
    eventos_match = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
        if score < 60:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
        eventos_match.append({
            'fecha': f.isoformat(),
            'titulo': ev.get('titulo'),
            'kg_parseado': kg,
            'score_match': score,
        })
        if f < fecha_pip and kg:
            kgs_hist.append(kg)

    cfg_lote = c.execute("""
        SELECT fh.lote_size_kg
        FROM formula_headers fh
        WHERE UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(?))
    """, (producto,)).fetchone()
    lote_default = (cfg_lote[0] if cfg_lote else None) or 30

    if kgs_hist:
        kgs_hist.sort()
        lote_tipico = kgs_hist[len(kgs_hist)//2]
    else:
        lote_tipico = lote_default

    out['lote'] = {
        'lote_default_formula': lote_default,
        'lote_tipico_historico': round(lote_tipico, 1),
        'historico_kg_lista': kgs_hist[:10],
        'eventos_calendar_match_count': len(eventos_match),
        'eventos_calendar_match_top10': eventos_match[:10],
    }

    # 8) Cálculo de unidades por lote
    out['unidades_por_lote'] = int((lote_tipico * 1000) / max(factor_g, 1))
    out['dias_que_durara_lote'] = round((out['unidades_por_lote'] / velocidad_proj), 1) if velocidad_proj > 0 else None

    # 9) Tabla de "ejemplo" para validar visualmente
    out['ejemplo_calculo'] = [
        f'1) Stock hoy = {stock_total} unidades (suma de {len(skus_map)} SKU{"s" if len(skus_map)!=1 else ""})',
        f'2) Velocidad = {round(velocidad_proj,2)} u/día (de ventas Shopify últimos 30d)',
        f'3) Días alcance = {out.get("dias_alcance_hoy", "N/A")} días',
        f'4) Margen mínimo = 20 días antes de agotar',
        f'5) Producir cuando alcance baje a margen',
        f'6) Lote = {round(lote_tipico,1)} kg ÷ {factor_g} g/u = {out["unidades_por_lote"]} unidades',
        f'7) Lote durará = {out.get("dias_que_durara_lote", "N/A")} días',
    ]

    # 10) Resumen de ÚLTIMOS pedidos Shopify (top 5) para verificar
    try:
        ultimos_pedidos = c.execute("""
            SELECT date(creado_en), nombre, unidades_total, sku_items
            FROM animus_shopify_orders
            ORDER BY creado_en DESC LIMIT 20
        """).fetchall()
        match_pedidos = []
        for fecha, nombre, _utot, sku_items_json in ultimos_pedidos:
            if not sku_items_json:
                continue
            try:
                items = json.loads(sku_items_json) if isinstance(sku_items_json, str) else sku_items_json
            except Exception:
                continue
            if not isinstance(items, list):
                continue
            for it in items:
                sk_pedido = (it.get('sku') or it.get('SKU') or '').strip()
                if any(sk_pedido == m[0] for m in skus_map):
                    match_pedidos.append({
                        'fecha': fecha,
                        'pedido': nombre,
                        'sku': sk_pedido,
                        'cantidad': it.get('cantidad') or it.get('quantity') or 0,
                    })
                    break
            if len(match_pedidos) >= 10:
                break
        out['ultimos_10_pedidos_con_este_producto'] = match_pedidos
    except Exception as e:
        out['ultimos_10_pedidos_con_este_producto'] = {'error': str(e)}

    return jsonify(out)


# ════════════════════════════════════════════════════════════════════════
# MP PARA LOTE — para cada producción del Calendar, ¿alcanza la materia prima?
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "ya sabemos hoy se hace este producto, qué
# necesita ese producto? Segundo, debe tener allí materias primas alcanza
# o no alcanza".

@bp.route('/api/planta/mp-para-lote', methods=['GET'])
def mp_para_lote():
    """Calcula MPs requeridas para un lote específico vs stock actual.

    Query params:
      - producto: nombre del producto (ej. 'SUERO HIDRATANTE AH 1.5%')
      - kg: tamaño del lote en kg

    Devuelve:
      {
        producto, lote_kg, total_g,
        mps: [{material_id, material_nombre, porcentaje, requerido_g,
               stock_g, falta_g, ratio, estado}],
        kpis: {total_mps, ok, ajustado, faltante, sin_formula},
        alcanza: bool (true si todas las MPs cubren ≥100% del requerimiento)
      }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    producto = (request.args.get('producto') or '').strip()
    if not producto:
        return jsonify({'error': 'Falta producto'}), 400
    try:
        kg = float(request.args.get('kg') or '0')
    except Exception:
        kg = 0
    if kg <= 0:
        return jsonify({'error': 'kg inválido'}), 400

    conn = get_db(); c = conn.cursor()

    # Buscar items de la fórmula
    items = c.execute("""
        SELECT material_id, material_nombre, porcentaje, COALESCE(cantidad_g_por_lote, 0)
        FROM formula_items
        WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
        ORDER BY porcentaje DESC
    """, (producto,)).fetchall()

    if not items:
        return jsonify({
            'producto': producto,
            'lote_kg': kg,
            'mps': [],
            'kpis': {'total_mps': 0, 'sin_formula': True},
            'alcanza': False,
            'mensaje': f'⚠️ Sin fórmula registrada para "{producto}". Ver Drive: Formulas Maestras',
        })

    # Stock map robusto: cruza 2 sistemas de IDs (formula vs bodega) + nombres
    # normalizados + alias + bridge table. Resuelve el bug de stock negativo
    # cuando formula_items.material_id no coincide con movimientos.material_id.
    try:
        from blueprints.programacion import _get_mp_stock, _norm_mp_name
        mp_stock_map = _get_mp_stock(conn)
    except Exception as e:
        log.warning(f'No pude cargar mp_stock_map robusto: {e}')
        mp_stock_map = None

    def _buscar_stock_robusto(mat_id, mat_nombre):
        """Busca stock en este orden:
           1. material_id directo (canónico bodega)
           2. material_id en uppercase
           3. nombre exacto en uppercase
           4. nombre normalizado (sin acentos, sin paréntesis, etc.)
           Devuelve (stock_g, fuente) - fuente para diagnóstico.
        """
        if mp_stock_map is None:
            # Fallback: query directa
            r = conn.execute("""
                SELECT COALESCE(SUM(
                    CASE WHEN tipo IN ('Entrada','Ingreso','Ajuste','Devolucion') THEN cantidad
                         WHEN tipo IN ('Salida','Consumo') THEN -cantidad
                         ELSE 0 END), 0) FROM movimientos WHERE material_id=?
            """, (mat_id,)).fetchone()
            return float(r[0] or 0) if r else 0, 'fallback_query'
        mid = str(mat_id or '').strip()
        if mid in mp_stock_map:
            return float(mp_stock_map[mid]), 'material_id'
        if mid.upper() in mp_stock_map:
            return float(mp_stock_map[mid.upper()]), 'material_id_upper'
        nom = str(mat_nombre or '').strip().upper()
        if nom and nom in mp_stock_map:
            return float(mp_stock_map[nom]), 'nombre_exacto'
        try:
            nom_norm = _norm_mp_name(mat_nombre or '')
            if nom_norm and nom_norm in mp_stock_map:
                return float(mp_stock_map[nom_norm]), 'nombre_normalizado'
        except Exception:
            pass
        return 0, 'no_encontrado'

    total_g_lote = kg * 1000
    mps = []
    falta_count = 0
    ajustado_count = 0
    ok_count = 0

    for material_id, mat_nombre, pct, g_ref in items:
        try:
            pct = float(pct or 0)
        except Exception:
            pct = 0
        req_g = total_g_lote * pct / 100

        # Stock con búsqueda robusta (resuelve diferencia de IDs)
        stock_g, fuente_stock = _buscar_stock_robusto(material_id, mat_nombre)

        # Lead time si está (mantenemos por material_id)
        lt = c.execute("""
            SELECT lead_time_dias, origen, proveedor_principal
            FROM mp_lead_time_config WHERE material_id=?
        """, (material_id,)).fetchone()
        lead_time = lt[0] if lt else None
        origen = lt[1] if lt else None
        proveedor = lt[2] if lt else None

        falta = req_g - stock_g
        ratio = stock_g / req_g if req_g > 0 else 99

        if ratio >= 1.10:
            estado = 'ok'
            ok_count += 1
        elif ratio >= 0.95:
            estado = 'ajustado'
            ajustado_count += 1
        else:
            estado = 'faltante'
            falta_count += 1

        mps.append({
            'material_id': material_id,
            'material_nombre': mat_nombre or material_id,
            'porcentaje': round(pct, 2),
            'requerido_g': round(req_g, 1),
            'stock_g': round(stock_g, 1),
            'falta_g': round(max(0, falta), 1),
            'ratio': round(ratio, 2),
            'estado': estado,
            'lead_time_dias': lead_time,
            'origen': origen,
            'proveedor': proveedor,
            'fuente_stock': fuente_stock,
        })

    alcanza = falta_count == 0

    # Mensaje resumen
    if alcanza and ajustado_count == 0:
        mensaje = f'✅ MP suficiente para los {kg} kg de "{producto}". {ok_count} ingredientes OK.'
    elif alcanza:
        mensaje = f'🟡 MP cubre el lote pero {ajustado_count} ingredientes ajustados (margen <10%).'
    else:
        # Lista críticos para mensaje
        criticos = [m['material_nombre'] for m in mps if m['estado'] == 'faltante'][:3]
        mensaje = f'🔴 FALTA MP para {falta_count} ingredientes' + ('. Críticos: ' + ', '.join(criticos) if criticos else '')

    return jsonify({
        'producto': producto,
        'lote_kg': kg,
        'total_g': total_g_lote,
        'mps': mps,
        'kpis': {
            'total_mps': len(mps),
            'ok': ok_count,
            'ajustado': ajustado_count,
            'faltante': falta_count,
            'sin_formula': False,
        },
        'alcanza': alcanza,
        'mensaje': mensaje,
        'fecha_analisis': datetime.now().isoformat(),
    })


# ════════════════════════════════════════════════════════════════════════
# ALERTAS CALENDAR — cruce de cadencia planeada vs velocidad real
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "ese calendario debería ir unido a una alerta
# que diga: se está vendiendo más, si está bien, o mejor adelanten".

@bp.route('/api/planta/alertas-calendar', methods=['GET'])
def alertas_calendar():
    """Para cada SKU con próximo lote en Calendar (próximos 60 días),
    compara la velocidad implícita en la cadencia (lo que asumimos al
    programar) vs la velocidad real medida en Shopify últimos 30 días.

    Devuelve alerta:
      - 🔴 ADELANTAR (vendes 20%+ más rápido)
      - 🟠 ADELANTAR_LIGERO (5-20% más)
      - 🟢 OK (±5%)
      - 🟡 ATRASAR_LIGERO (15-35% menos)
      - ⚠️ REDUCIR_LOTE (35%+ menos vendes)
      - ❓ SIN_DATOS (no hay velocidad real medible)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()

    # Leer todos los eventos del Calendar (próximos 60 días)
    eventos = _calendar_events_cached()

    # Para cada producto activo, calcular alerta
    skus = c.execute("""
        SELECT spc.producto_nombre, spc.cadencia_dias, fh.lote_size_kg,
               COALESCE(spc.estado, 'activo')
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
          AND COALESCE(spc.estado, 'activo') NOT IN ('descontinuado', 'pausado')
        ORDER BY spc.prioridad ASC
    """).fetchall()

    alertas = []
    for sku_row in skus:
        producto, cadencia_cfg, lote_kg, estado = sku_row
        alias = _alias_calendar_for(c, producto)

        # Buscar próxima fecha programada en Calendar (futuro <= 60 días)
        proxima_fecha = None
        proximo_titulo = None
        kg_proximo = None
        eventos_futuros = []
        for ev in eventos:
            try:
                f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            except Exception:
                continue
            if f < fecha_hoy or (f - fecha_hoy).days > 60:
                continue
            score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
            if score < 60:
                continue
            kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
            eventos_futuros.append({'fecha': f, 'titulo': ev.get('titulo'), 'kg': kg})

        if not eventos_futuros:
            # Sin lote programado en próximos 60 días
            continue

        eventos_futuros.sort(key=lambda x: x['fecha'])
        proxima_fecha = eventos_futuros[0]['fecha']
        proximo_titulo = eventos_futuros[0]['titulo']
        kg_proximo = eventos_futuros[0]['kg'] or lote_kg or 30

        # Cadencia entre próximo lote y siguiente (si hay) → si no, usar cadencia configurada
        if len(eventos_futuros) >= 2:
            cadencia_real_cal = (eventos_futuros[1]['fecha'] - eventos_futuros[0]['fecha']).days
        else:
            cadencia_real_cal = cadencia_cfg or 60

        # Velocidad implícita: cuánto necesitarías vender para que el lote aguante esa cadencia
        factor_g = _factor_g_por_unidad(c, producto)
        unidades_lote = (kg_proximo * 1000) / max(factor_g, 1)
        velocidad_planeada = unidades_lote / max(cadencia_real_cal + 20, 30)  # con margen 20d

        # Velocidad real
        velocidad_real, factor_tend = _velocidad_total_producto(c, producto)
        velocidad_real_proj = velocidad_real * factor_tend if factor_tend > 0 else velocidad_real

        # Si velocidad real prácticamente cero, no es alerta accionable
        if velocidad_real_proj < 0.05 and velocidad_planeada > 0.5:
            alertas.append({
                'producto': producto,
                'proxima_fecha': proxima_fecha.isoformat(),
                'dias_hasta_proximo': (proxima_fecha - fecha_hoy).days,
                'kg_proximo': kg_proximo,
                'cadencia_calendar_dias': cadencia_real_cal,
                'velocidad_planeada': round(velocidad_planeada, 2),
                'velocidad_real': round(velocidad_real_proj, 2),
                'ratio': 0,
                'estado': 'sin_ventas',
                'mensaje': 'Sin ventas detectadas. Considerar pausar o evaluar lote.',
                'titulo_evento': proximo_titulo,
            })
            continue

        ratio = velocidad_real_proj / max(velocidad_planeada, 0.01)

        # Calcular días de adelanto/atraso
        # Nuevo alcance del próximo lote = unidades_lote / velocidad_real
        nuevo_alcance = unidades_lote / max(velocidad_real_proj, 0.01)
        # Idealmente lote dura cadencia + 20d margen. Diferencia = días adelantar/atrasar
        ideal_dura = cadencia_real_cal + 20
        diff_dias = round(ideal_dura - nuevo_alcance)  # positivo: hay que adelantar

        if ratio >= 1.20:
            estado = 'adelantar'
            mensaje = f'🔴 Vendes {int((ratio-1)*100)}% más rápido. Adelantar lote ~{abs(diff_dias)}d'
            urg = 'alta'
        elif ratio >= 1.05:
            estado = 'adelantar_ligero'
            mensaje = f'🟠 Vendes {int((ratio-1)*100)}% más. Considera adelantar ~{abs(diff_dias)}d'
            urg = 'media'
        elif ratio >= 0.85:
            estado = 'ok'
            mensaje = f'🟢 Velocidad coincide con plan ({int((ratio-1)*100):+d}%). Sigue como está.'
            urg = 'ok'
        elif ratio >= 0.65:
            estado = 'atrasar_ligero'
            mensaje = f'🟡 Vendes {int((1-ratio)*100)}% menos. Considera atrasar lote ~{abs(diff_dias)}d'
            urg = 'media'
        else:
            estado = 'reducir_lote'
            mensaje = f'⚠️ Vendes {int((1-ratio)*100)}% menos. Evaluar reducir kg del lote o atrasar mucho'
            urg = 'media'

        alertas.append({
            'producto': producto,
            'proxima_fecha': proxima_fecha.isoformat(),
            'dias_hasta_proximo': (proxima_fecha - fecha_hoy).days,
            'kg_proximo': round(kg_proximo, 1),
            'unidades_lote': int(unidades_lote),
            'cadencia_calendar_dias': cadencia_real_cal,
            'velocidad_planeada': round(velocidad_planeada, 2),
            'velocidad_real': round(velocidad_real_proj, 2),
            'ratio': round(ratio, 2),
            'diff_dias': diff_dias,
            'estado': estado,
            'urgencia': urg,
            'mensaje': mensaje,
            'titulo_evento': proximo_titulo,
            'nuevo_alcance_dias': round(nuevo_alcance),
        })

    # Ordenar: las más urgentes primero
    orden_estado = {'adelantar': 0, 'adelantar_ligero': 1, 'reducir_lote': 2,
                    'atrasar_ligero': 3, 'sin_ventas': 4, 'ok': 5}
    alertas.sort(key=lambda a: (orden_estado.get(a['estado'], 9), a.get('dias_hasta_proximo', 999)))

    # KPIs
    kpis = {
        'total': len(alertas),
        'adelantar': sum(1 for a in alertas if a['estado'] == 'adelantar'),
        'adelantar_ligero': sum(1 for a in alertas if a['estado'] == 'adelantar_ligero'),
        'ok': sum(1 for a in alertas if a['estado'] == 'ok'),
        'atrasar_ligero': sum(1 for a in alertas if a['estado'] == 'atrasar_ligero'),
        'reducir_lote': sum(1 for a in alertas if a['estado'] == 'reducir_lote'),
        'sin_ventas': sum(1 for a in alertas if a['estado'] == 'sin_ventas'),
    }

    return jsonify({
        'alertas': alertas,
        'kpis': kpis,
        'fecha_analisis': datetime.now().isoformat(),
        'reglas': [
            'Velocidad real = ventas Shopify últimos 30d',
            'Velocidad planeada = unidades_lote / (cadencia_calendar + 20d margen)',
            'Ratio ≥ 1.20 → 🔴 ADELANTAR (vendes 20%+ rápido)',
            'Ratio 1.05-1.20 → 🟠 ADELANTAR LIGERO',
            'Ratio 0.85-1.05 → 🟢 OK',
            'Ratio 0.65-0.85 → 🟡 ATRASAR LIGERO',
            'Ratio < 0.65 → ⚠️ REDUCIR LOTE',
        ],
    })


# ════════════════════════════════════════════════════════════════════════
# PLAN PURO SHOPIFY — vista por día (próximo lunes → viernes siguiente)
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "no tengas en cuenta el calendario, si tuvieras
# que producir ya, según Shopify inventario y ventas dime cómo lo harías
# aquí, qué iría el próximo lunes, martes, miércoles etc".

@bp.route('/api/planta/plan-semana-shopify', methods=['GET'])
def plan_semana_shopify():
    """Plan de la próxima semana basado SOLO en Shopify (stock + ventas), ignorando Calendar.

    Devuelve:
      {
        'semana_inicio': 'YYYY-MM-DD' (próximo lunes),
        'semana_fin':    'YYYY-MM-DD' (viernes siguiente),
        'dias': [
          {
            'fecha': 'YYYY-MM-DD', 'nombre_dia': 'lunes', 'orden_dia': 0,
            'producciones': [
              {producto, lote_kg, velocidad_dia, stock_actual, dias_alcance,
               urgencia, razon, factor_g, dias_que_durara_lote}
            ]
          }, ...
        ],
        'sin_fecha': [...],   // recomendaciones que caen >7d (no esta semana)
        'kpis': {...}
      }

    Query params:
      - semanas: int (default 1) → cuántas semanas planear (1, 2, 4)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    semanas = max(1, min(8, int(request.args.get('semanas', '1'))))
    conn = get_db(); c = conn.cursor()

    skus = c.execute("""
        SELECT spc.producto_nombre, spc.cadencia_dias, spc.merma_pct, spc.prioridad,
               fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
        ORDER BY spc.prioridad ASC
    """).fetchall()

    # Calcular recomendaciones IGNORANDO Calendar
    recs = []
    for sku_row in skus:
        producto, cadencia, merma, prioridad, lote_kg = sku_row
        try:
            r = _calcular_recomendacion_sku(c, producto, lote_kg, cadencia, merma, prioridad, ignorar_calendar=True)
            recs.append(r)
        except Exception as e:
            log.warning(f'plan-semana-shopify recomendación fallida {producto}: {e}')

    # Solo accionables (urgencia que requiere producir)
    accionables = [r for r in recs if r.get('urgencia') in ('critica', 'alta', 'media', 'baja')]

    # Distribución: si >3 accionables/semana → L-V, si <=3 → L/M/V
    distribuir_lv_completo = len(accionables) > 3 * semanas
    DIAS_PRODUCCION = (0, 1, 2, 3, 4) if distribuir_lv_completo else (0, 2, 4)

    fecha_hoy = datetime.now().date()
    # Próximo lunes (si hoy es lunes y son antes de las 12, hoy mismo; si no, lunes siguiente)
    dias_a_lunes = (7 - fecha_hoy.weekday()) % 7
    if dias_a_lunes == 0:
        # Hoy es lunes
        proximo_lunes = fecha_hoy
    else:
        proximo_lunes = fecha_hoy + timedelta(days=dias_a_lunes)

    semana_fin = proximo_lunes + timedelta(days=(7 * semanas) - 3)  # Viernes de la última semana

    # Construir slots disponibles en la(s) semana(s)
    slots = []
    f = proximo_lunes
    for _ in range(7 * semanas):
        if f.weekday() in DIAS_PRODUCCION:
            slots.append(f)
        f += timedelta(days=1)

    # Ordenar accionables por urgencia + días alcance ascendente (más urgente primero)
    orden_urg = {'critica': 0, 'alta': 1, 'media': 2, 'baja': 3}
    accionables.sort(key=lambda r: (orden_urg.get(r['urgencia'], 9), float(r.get('dias_alcance') or 9999)))

    # Asignar 1 producción por slot (regla: 1 producto/día por área dispensación)
    LIMITE_POR_DIA = 1
    asignados_por_dia = {s.isoformat(): [] for s in slots}
    sin_slot = []

    for r in accionables:
        # Días hasta que sea necesario producir (margen 20d)
        dias_hasta = max(0, int(r.get('dias_alcance', 30)) - 20)
        fecha_ideal = fecha_hoy + timedelta(days=dias_hasta)

        # Buscar primer slot >= fecha_ideal con cupo
        asignado = False
        for s in slots:
            if s < fecha_ideal:
                continue
            if len(asignados_por_dia[s.isoformat()]) < LIMITE_POR_DIA:
                # Calcular cuánto durará el lote producido (para decidir cadencia)
                lote_kg_use = r.get('lote_tipico_kg') or r.get('lote_kg_default') or 30
                factor_g = r.get('factor_g') or 30
                velocidad = max(0.01, r.get('velocidad_dia') or 0.01)
                unidades_lote = (lote_kg_use * 1000) / max(factor_g, 1)
                dias_que_durara = int(unidades_lote / velocidad)
                asignados_por_dia[s.isoformat()].append({
                    'producto': r['producto'],
                    'lote_kg': round(lote_kg_use, 1),
                    'velocidad_dia': r['velocidad_dia'],
                    'stock_actual': r.get('stock_shopify', 0),
                    'dias_alcance': r['dias_alcance'],
                    'urgencia': r['urgencia'],
                    'razon': r['razon'],
                    'factor_g': factor_g,
                    'dias_que_durara_lote': dias_que_durara,
                    'unidades_lote': int(unidades_lote),
                })
                asignado = True
                break
        if not asignado:
            # No cupo en la semana → fuera
            sin_slot.append({
                'producto': r['producto'],
                'lote_kg': round(r.get('lote_tipico_kg') or r.get('lote_kg_default') or 30, 1),
                'dias_alcance': r['dias_alcance'],
                'urgencia': r['urgencia'],
                'razon': 'No hay cupo en la(s) semana(s) planeada(s). Programar después.',
                'fecha_proxima_estimada': r.get('fecha_proxima'),
            })

    # Construir respuesta por día
    nombres_dia = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']
    dias_out = []
    f = proximo_lunes
    for _ in range(7 * semanas):
        es_dia_prod = f.weekday() in DIAS_PRODUCCION
        producciones = asignados_por_dia.get(f.isoformat(), [])
        # Solo incluir días con producciones o días de producción aunque vacíos
        if es_dia_prod:
            dias_out.append({
                'fecha': f.isoformat(),
                'nombre_dia': nombres_dia[f.weekday()],
                'orden_dia': f.weekday(),
                'es_dia_produccion': True,
                'producciones': producciones,
                'producciones_count': len(producciones),
            })
        f += timedelta(days=1)

    # KPIs
    kpis = {
        'total_recs': len(recs),
        'accionables': len(accionables),
        'criticas': sum(1 for r in accionables if r['urgencia'] == 'critica'),
        'altas': sum(1 for r in accionables if r['urgencia'] == 'alta'),
        'medias': sum(1 for r in accionables if r['urgencia'] == 'media'),
        'bajas': sum(1 for r in accionables if r['urgencia'] == 'baja'),
        'asignadas_semana': sum(len(d['producciones']) for d in dias_out),
        'sin_cupo': len(sin_slot),
        'slots_disponibles': len(slots),
    }

    return jsonify({
        'semana_inicio': proximo_lunes.isoformat(),
        'semana_fin': semana_fin.isoformat(),
        'semanas': semanas,
        'patron_distribucion': 'L-V (5 días)' if distribuir_lv_completo else 'L/M/V (3 días)',
        'dias': dias_out,
        'sin_slot': sin_slot,
        'kpis': kpis,
        'fecha_analisis': datetime.now().isoformat(),
        'modo': 'shopify_puro_sin_calendar',
        'reglas': [
            'Stock = Shopify (NO suma Calendar pipeline ni futuro)',
            'Velocidad = ventas Shopify últimos 30d',
            'Margen mínimo = 20d antes de agotar',
            f'Distribución = {"L-V (5 días)" if distribuir_lv_completo else "L/M/V (3 días)"}',
            f'Límite = {LIMITE_POR_DIA} producción/día (área dispensación = Mayerlin)',
            'Orden = críticas primero, luego por días-alcance',
        ],
    })


# ════════════════════════════════════════════════════════════════════════
# PLAN LARGO SHOPIFY — rolling forecast día-a-día (6 meses, 1 año)
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "quiero que automáticamente la app hoy reconozca
# inventario en shopy cruce con ventas y sepa como produciremos 6 meses y
# 1 año, es lo más importante. Lo primero es que el sistema reconozca qué
# producir".

@bp.route('/api/planta/plan-largo-shopify', methods=['GET'])
def plan_largo_shopify():
    """Rolling forecast SHOPIFY-only para 6 o 12 meses.

    Para cada SKU activo simula día-a-día:
      stock_dia = stock_inicial - velocidad * dias_transcurridos
      Cuando stock_dia <= velocidad * 20 (margen) → programar producción
        en el próximo L/M/V con cupo libre (1 SKU/día regla Mayerlin).
      Tras producir: stock_dia += unidades_lote (efectivo al día siguiente).
      Repetir hasta horizonte_dias.

    Devuelve:
      {
        'horizonte_meses': 6 o 12,
        'fecha_inicio': hoy,
        'fecha_fin': hoy + N días,
        'producciones': [
          {fecha, producto, lote_kg, unidades_lote, motivo,
           stock_antes, stock_despues, dia_semana, mes}
        ],
        'producciones_por_mes': {
          '2026-05': [{producto, lote_kg, fecha}],
          '2026-06': [...]
        },
        'producciones_por_sku': {
          'Producto X': {
            'total_lotes': 8, 'total_kg': 720,
            'fechas': ['2026-05-04','2026-05-30',...]
          }
        },
        'kpis': {
          total_lotes, total_kg, productos_planeados,
          dias_con_produccion, slots_libres,
          alerta_capacidad: bool (si hay >1 SKU/día forzado)
        },
        'sin_ventas': [...] // SKUs descartados por velocidad ~0
      }

    Query params:
      - meses: 6 o 12 (default 6)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    meses = max(1, min(24, int(request.args.get('meses', '6'))))
    horizonte_dias = meses * 30
    conn = get_db(); c = conn.cursor()

    skus = c.execute("""
        SELECT spc.producto_nombre, spc.cadencia_dias, spc.merma_pct, spc.prioridad,
               fh.lote_size_kg, spc.estado
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
          AND COALESCE(spc.estado,'activo') NOT IN ('descontinuado','pausado')
        ORDER BY spc.prioridad ASC
    """).fetchall()

    fecha_hoy = datetime.now().date()
    margen_dias = 20  # Sebastián: producir 20d antes de agotar

    # Reunir info por SKU primero
    sku_info = []
    sin_ventas = []
    for sku_row in skus:
        producto, cadencia_cfg, _merma, prioridad, lote_kg_default, estado = sku_row
        velocidad, factor = _velocidad_total_producto(c, producto)
        velocidad_proj = max(0.0, velocidad * factor)
        stock_inicial = _stock_actual_pt(c, producto)
        factor_g = _factor_g_por_unidad(c, producto)

        # Lote típico desde histórico (mediana) o default
        eventos_hist = _calendar_events_cached()
        alias = _alias_calendar_for(c, producto)
        kg_hist = []
        for ev in eventos_hist:
            try:
                f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            except Exception:
                continue
            if f >= fecha_hoy - timedelta(days=14):
                continue
            score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
            if score < 60:
                continue
            kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
            if kg:
                kg_hist.append(kg)
        if kg_hist:
            kg_hist.sort()
            lote_kg_use = kg_hist[len(kg_hist)//2]
        else:
            lote_kg_use = lote_kg_default or 30

        unidades_lote = (lote_kg_use * 1000) / max(factor_g, 1)

        # Si velocidad ~0 → no se planea
        if velocidad_proj < 0.01:
            sin_ventas.append({
                'producto': producto,
                'razon': 'Sin ventas detectadas en Shopify (velocidad < 0.01 u/día)',
                'stock_inicial': int(stock_inicial),
            })
            continue

        sku_info.append({
            'producto': producto,
            'velocidad': velocidad_proj,
            'stock_inicial': stock_inicial,
            'factor_g': factor_g,
            'lote_kg': lote_kg_use,
            'unidades_lote': unidades_lote,
            'prioridad': prioridad or 50,
        })

    # ROLLING FORECAST: simular cada día del horizonte y programar lotes
    # Estrategia: para cada SKU calcular las fechas en las que stock toca margen,
    # luego asignar cada una al primer L/M/V con cupo libre.
    LIMITE_POR_DIA = 1
    cupo_por_dia = {}  # iso_fecha -> count

    todas_producciones = []  # lista plana de {fecha, producto, lote_kg, ...}

    # 1) Para cada SKU: identifica fechas-objetivo donde necesita lote
    for s in sku_info:
        velocidad = s['velocidad']
        unidades_lote = s['unidades_lote']
        stock = s['stock_inicial']
        cursor_dia = 0  # días transcurridos desde hoy
        n_lotes_sku = 0
        max_lotes_sku = max(1, int((velocidad * horizonte_dias) / unidades_lote) + 2)

        while cursor_dia <= horizonte_dias and n_lotes_sku < max_lotes_sku:
            # ¿Cuántos días hasta que stock baje al margen?
            dias_hasta_margen = max(0, int((stock - velocidad * margen_dias) / velocidad))
            fecha_objetivo = fecha_hoy + timedelta(days=cursor_dia + dias_hasta_margen)
            # Si la fecha ya pasó del horizonte, fin
            if (fecha_objetivo - fecha_hoy).days > horizonte_dias:
                break
            # Buscar primer L/M/V con cupo desde fecha_objetivo
            f_busqueda = max(fecha_objetivo, fecha_hoy + timedelta(days=2))
            asignado = None
            DIAS_VALIDOS_FALLBACK = (0, 1, 2, 3, 4)  # L-V completos
            for _ in range(60):
                if f_busqueda.weekday() in (0, 2, 4):  # L M V primero
                    iso = f_busqueda.isoformat()
                    if cupo_por_dia.get(iso, 0) < LIMITE_POR_DIA:
                        asignado = f_busqueda
                        break
                f_busqueda += timedelta(days=1)
            # Si no cupo en L/M/V → probar Ma/Ju
            if asignado is None:
                f_busqueda = max(fecha_objetivo, fecha_hoy + timedelta(days=2))
                for _ in range(60):
                    if f_busqueda.weekday() in DIAS_VALIDOS_FALLBACK:
                        iso = f_busqueda.isoformat()
                        if cupo_por_dia.get(iso, 0) < LIMITE_POR_DIA:
                            asignado = f_busqueda
                            break
                    f_busqueda += timedelta(days=1)
            if asignado is None:
                # Forzar (alerta de capacidad)
                asignado = max(fecha_objetivo, fecha_hoy + timedelta(days=2))
                # Avanzar al primer día laborable
                while asignado.weekday() not in DIAS_VALIDOS_FALLBACK:
                    asignado += timedelta(days=1)

            # Calcular stock antes (en fecha asignada) y después
            dias_real = (asignado - fecha_hoy).days
            stock_antes = max(0, stock - velocidad * (dias_real - cursor_dia))
            stock_despues = stock_antes + unidades_lote

            iso_asignado = asignado.isoformat()
            cupo_por_dia[iso_asignado] = cupo_por_dia.get(iso_asignado, 0) + 1

            todas_producciones.append({
                'fecha': iso_asignado,
                'producto': s['producto'],
                'lote_kg': round(s['lote_kg'], 1),
                'unidades_lote': int(unidades_lote),
                'stock_antes': int(stock_antes),
                'stock_despues': int(stock_despues),
                'velocidad_dia': round(velocidad, 2),
                'factor_g': s['factor_g'],
                'dia_semana': ['lunes','martes','miércoles','jueves','viernes','sábado','domingo'][asignado.weekday()],
                'mes': asignado.strftime('%Y-%m'),
                'forzado_capacidad': cupo_por_dia.get(iso_asignado, 0) > LIMITE_POR_DIA,
                'motivo': f'Stock cae a margen 20d el {(fecha_hoy + timedelta(days=cursor_dia + dias_hasta_margen)).isoformat()}',
            })

            # Avanzar simulación: stock se reabastece, cursor avanza dias_real
            stock = stock_despues
            cursor_dia = dias_real + 1
            n_lotes_sku += 1

    # Ordenar por fecha
    todas_producciones.sort(key=lambda p: (p['fecha'], p['producto']))

    # Agrupar por mes
    por_mes = {}
    for p in todas_producciones:
        por_mes.setdefault(p['mes'], []).append({
            'producto': p['producto'],
            'lote_kg': p['lote_kg'],
            'fecha': p['fecha'],
            'dia_semana': p['dia_semana'],
        })

    # Agrupar por SKU
    por_sku = {}
    for p in todas_producciones:
        if p['producto'] not in por_sku:
            por_sku[p['producto']] = {
                'total_lotes': 0,
                'total_kg': 0,
                'fechas': [],
                'velocidad_dia': p['velocidad_dia'],
                'factor_g': p['factor_g'],
            }
        por_sku[p['producto']]['total_lotes'] += 1
        por_sku[p['producto']]['total_kg'] += p['lote_kg']
        por_sku[p['producto']]['fechas'].append(p['fecha'])

    # KPIs
    dias_unicos = set(p['fecha'] for p in todas_producciones)
    forzados = sum(1 for p in todas_producciones if p.get('forzado_capacidad'))
    kpis = {
        'total_lotes': len(todas_producciones),
        'total_kg': round(sum(p['lote_kg'] for p in todas_producciones), 1),
        'productos_planeados': len(por_sku),
        'productos_sin_ventas': len(sin_ventas),
        'dias_con_produccion': len(dias_unicos),
        'meses_cubiertos': len(por_mes),
        'forzados_por_capacidad': forzados,
        'alerta_capacidad': forzados > 0,
        'promedio_lotes_por_mes': round(len(todas_producciones) / max(meses, 1), 1),
    }

    return jsonify({
        'horizonte_meses': meses,
        'horizonte_dias': horizonte_dias,
        'fecha_inicio': fecha_hoy.isoformat(),
        'fecha_fin': (fecha_hoy + timedelta(days=horizonte_dias)).isoformat(),
        'producciones': todas_producciones,
        'producciones_por_mes': por_mes,
        'producciones_por_sku': por_sku,
        'sin_ventas': sin_ventas,
        'kpis': kpis,
        'fecha_analisis': datetime.now().isoformat(),
        'modo': 'shopify_rolling_forecast',
        'reglas': [
            f'Horizonte = {meses} meses ({horizonte_dias} días)',
            'Stock inicial = Shopify · Velocidad = ventas Shopify 30d',
            'Programar lote cuando stock cae a margen 20d',
            'Tras producir: stock += unidades_lote (al día siguiente)',
            'Asignación: L/M/V primero, luego Ma/Ju, 1 SKU/día (Mayerlin)',
            'Lote típico = mediana histórica Calendar o default',
        ],
    })


# ════════════════════════════════════════════════════════════════════════
# AUDITORÍA del Calendar — ¿se cumplió margen 20d?
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "la pregunta real es lo que está en el calendario
# si cumple la lógica de producto planeado 20 días antes de que se agote".

@bp.route('/api/planta/auditoria-calendar', methods=['GET'])
def auditoria_calendar():
    """Para cada producción en el Calendar, evalúa si se hizo respetando el
    margen de 20 días antes de agotar stock.

    Lógica:
      Para cada lote del calendar histórico:
        - Calcula duración estimada del lote = kg × 1000 / factor_g / velocidad
        - Calcula gap real entre esta producción y la anterior
        - margen_real = duración_anterior - gap
        - Si margen >= 20d   → OK ✓
        - Si margen 5-20d    → AJUSTADA (debió producirse antes)
        - Si margen 0-5d     → TARDE (al límite)
        - Si margen < 0      → STOCK-OUT (producción tarde, hubo días sin stock)
        - Si margen > 40d    → TEMPRANA (overstock)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    eventos = _calendar_events_cached()
    productos_cfg = c.execute("""
        SELECT producto_nombre, alias_calendar FROM sku_planeacion_config WHERE activo=1
    """).fetchall()

    # Agrupar eventos por producto
    eventos_por_producto = {}  # producto → [(fecha, kg)]
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        mejor_score = 0
        mejor_producto = None
        for p in productos_cfg:
            score = _match_producto_evento(p[0], p[1], ev.get('titulo'), ev.get('descripcion', ''))
            if score > mejor_score:
                mejor_score = score
                mejor_producto = p[0]
        if mejor_score < 60 or not mejor_producto:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
        eventos_por_producto.setdefault(mejor_producto, []).append({
            'fecha': f, 'kg': kg, 'titulo': ev.get('titulo', '')
        })

    # Auditar cada lote
    auditorias = []
    for producto, lotes in eventos_por_producto.items():
        if len(lotes) < 2:
            continue
        lotes.sort(key=lambda x: x['fecha'])
        velocidad, factor = _velocidad_total_producto(c, producto)
        velocidad_proj = max(0.01, velocidad * factor)
        factor_g = _factor_g_por_unidad(c, producto)

        for i in range(1, len(lotes)):
            lote_actual = lotes[i]
            lote_anterior = lotes[i-1]
            gap_dias = (lote_actual['fecha'] - lote_anterior['fecha']).days
            kg_anterior = lote_anterior['kg'] or 30
            unidades_anterior = (kg_anterior * 1000) / max(factor_g, 1)
            duracion_lote_anterior = unidades_anterior / max(velocidad_proj, 0.01)
            margen_real = duracion_lote_anterior - gap_dias

            if margen_real >= 40:
                clase = 'temprana'
                msg = f'Lote anterior ({kg_anterior}kg) cubría {duracion_lote_anterior:.0f}d, gap {gap_dias}d → produjiste {margen_real:.0f}d antes (overstock)'
            elif margen_real >= 20:
                clase = 'ok'
                msg = f'Lote anterior cubría {duracion_lote_anterior:.0f}d, gap {gap_dias}d → margen {margen_real:.0f}d ≥ 20d ✓'
            elif margen_real >= 5:
                clase = 'ajustada'
                msg = f'Margen {margen_real:.0f}d (debajo de 20d ideal pero positivo)'
            elif margen_real >= 0:
                clase = 'tarde'
                msg = f'Margen {margen_real:.0f}d — producción al límite'
            else:
                clase = 'stockout'
                msg = f'Stock se agotó {abs(margen_real):.0f}d antes de la nueva producción'

            auditorias.append({
                'producto': producto,
                'fecha_actual': lote_actual['fecha'].isoformat(),
                'fecha_anterior': lote_anterior['fecha'].isoformat(),
                'kg_anterior': kg_anterior,
                'gap_dias': gap_dias,
                'duracion_estimada_lote': round(duracion_lote_anterior, 1),
                'margen_dias': round(margen_real, 1),
                'velocidad_dia': round(velocidad_proj, 2),
                'clase': clase,
                'mensaje': msg,
            })

    # Resumen
    total = len(auditorias)
    by_clase = {}
    for a in auditorias:
        by_clase[a['clase']] = by_clase.get(a['clase'], 0) + 1

    cumple_margen = (by_clase.get('ok', 0) + by_clase.get('temprana', 0)) / max(total, 1) * 100

    return jsonify({
        'auditorias': auditorias,
        'total': total,
        'kpis': {
            'cumple_margen_pct': round(cumple_margen, 1),
            'ok': by_clase.get('ok', 0),
            'temprana': by_clase.get('temprana', 0),
            'ajustada': by_clase.get('ajustada', 0),
            'tarde': by_clase.get('tarde', 0),
            'stockout': by_clase.get('stockout', 0),
        },
        'productos_evaluados': len(eventos_por_producto),
    })


@bp.route('/api/planta/calendar-eventos-plan', methods=['GET'])
def calendar_eventos_plan():
    """Devuelve eventos del Google Calendar que el motor MRP matchea con
    productos configurados, para mostrarlos EN el calendario del Plan v2.

    Sebastian (30-abr-2026): "tal cual como es" — el calendario real debe
    aparecer en el Plan, no solo las proyecciones del motor.

    Querystring: ?dias=30|60|90|180|365 (default 30)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 30))
    except Exception:
        dias = 30
    conn = get_db(); c = conn.cursor()
    eventos = _calendar_events_cached()
    fecha_hoy = datetime.now().date()
    fecha_top = fecha_hoy + timedelta(days=dias)

    productos = c.execute("""
        SELECT producto_nombre, alias_calendar FROM sku_planeacion_config WHERE activo=1
    """).fetchall()
    productos_list = [{'nombre': p[0], 'alias': p[1]} for p in productos]

    items = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if f < fecha_hoy or f > fecha_top:
            continue
        # Match
        mejor_score = 0
        mejor_producto = None
        for p in productos_list:
            score = _match_producto_evento(p['nombre'], p['alias'], ev.get('titulo'), ev.get('descripcion', ''))
            if score > mejor_score:
                mejor_score = score
                mejor_producto = p['nombre']
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
        items.append({
            'fecha': f.isoformat(),
            'titulo': ev.get('titulo', ''),
            'producto_match': mejor_producto if mejor_score >= 60 else None,
            'score': mejor_score,
            'kg': kg,
            'origen': 'google_calendar',
        })
    items.sort(key=lambda x: x['fecha'])
    return jsonify({'eventos': items, 'total': len(items)})


@bp.route('/api/planta/calendar-debug', methods=['GET'])
def calendar_debug():
    """Devuelve TODOS los eventos del calendar, su match con productos
    configurados, y kg detectados. Sebastian: "que sea perfecto"."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    eventos = _calendar_events_cached(force_refresh=True)
    productos = c.execute("""
        SELECT spc.producto_nombre, spc.alias_calendar, fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre))=UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo=1
    """).fetchall()
    productos_list = [{'nombre': p[0], 'alias': p[1], 'lote_kg': p[2]} for p in productos]
    eventos_analizados = []
    for ev in eventos:
        # Buscar mejor match
        mejor_score = 0
        mejor_producto = None
        candidatos = []
        for p in productos_list:
            score = _match_producto_evento(p['nombre'], p['alias'], ev.get('titulo'), ev.get('descripcion', ''))
            if score > 0:
                candidatos.append({'producto': p['nombre'], 'score': score})
                if score > mejor_score:
                    mejor_score = score
                    mejor_producto = p['nombre']
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
        candidatos.sort(key=lambda x: -x['score'])
        # Estado
        if mejor_score >= 60:
            if len([cc for cc in candidatos if cc['score'] >= 60]) > 1:
                estado = 'conflicto'
            else:
                estado = 'matcheado'
        elif mejor_score > 0:
            estado = 'sin_match'
        else:
            estado = 'no_relacionado'
        eventos_analizados.append({
            'titulo': ev.get('titulo', ''),
            'fecha': ev.get('fecha', ''),
            'descripcion': (ev.get('descripcion') or '')[:200],
            'kg_detectados': kg,
            'producto_match': mejor_producto if mejor_score >= 60 else None,
            'score_match': mejor_score,
            'candidatos_top3': candidatos[:3],
            'estado': estado,
        })
    # KPIs
    total = len(eventos_analizados)
    matcheados = sum(1 for e in eventos_analizados if e['estado'] == 'matcheado')
    conflicto = sum(1 for e in eventos_analizados if e['estado'] == 'conflicto')
    sin_match = sum(1 for e in eventos_analizados if e['estado'] == 'sin_match')
    no_rel = sum(1 for e in eventos_analizados if e['estado'] == 'no_relacionado')
    con_kg = sum(1 for e in eventos_analizados if e['kg_detectados'])
    return jsonify({
        'total_eventos': total,
        'matcheados': matcheados,
        'en_conflicto': conflicto,
        'sin_match_aceptable': sin_match,
        'no_relacionados': no_rel,
        'con_kg_detectados': con_kg,
        'eventos': eventos_analizados,
        'productos_configurados': len(productos_list),
    })


# ════════════════════════════════════════════════════════════════════════
# APRENDIZAJE DEL HISTÓRICO — el sistema infiere cadencias reales
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "en el calendario aparece lo que ya fabricamos
# es decir esos podrías usarlo como universo para el futuro y arrancar
# colocando lo que no hemos producido".
#
# Lee TODAS las producciones histórias (produccion_programada estado
# completado/en_proceso + Google Calendar) y deriva:
#   - Cadencia real (mediana de intervalos entre lotes)
#   - Última fecha de producción
#   - Productos NUEVOS sin histórico → arrancan como "primer lote"

def _calcular_cadencia_real(fechas):
    """Dada una lista de fechas (date objects), calcula intervalos en días
    y devuelve la mediana."""
    if len(fechas) < 2:
        return None
    fechas_ordenadas = sorted(fechas)
    intervalos = []
    for i in range(1, len(fechas_ordenadas)):
        delta = (fechas_ordenadas[i] - fechas_ordenadas[i-1]).days
        if delta > 0:
            intervalos.append(delta)
    if not intervalos:
        return None
    intervalos.sort()
    n = len(intervalos)
    if n % 2 == 0:
        return (intervalos[n//2 - 1] + intervalos[n//2]) // 2
    return intervalos[n // 2]


@bp.route('/api/planta/kpi-cobertura', methods=['GET'])
def kpi_cobertura_skus():
    """Sebastian (30-abr-2026): "dime cuántos productos están en el plan
    para saber que sí están todos los SKU".

    Querystring: ?dias=14|30|60|90|180|365 — calcula cobertura sobre el
    horizonte solicitado. La cobertura considera producción en BD +
    Google Calendar.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 60))
    except Exception:
        dias = 60
    conn = get_db(); c = conn.cursor()
    productos_total = c.execute(
        "SELECT producto_nombre FROM formula_headers WHERE producto_nombre IS NOT NULL"
    ).fetchall()
    productos_total_set = {(p[0] or '').strip().upper() for p in productos_total if p[0]}

    fecha_limite = (datetime.now() + timedelta(days=dias)).strftime('%Y-%m-%d')

    # Producciones en BD dentro del horizonte
    productos_en_plan_bd = c.execute("""
        SELECT DISTINCT UPPER(TRIM(producto)) FROM produccion_programada
        WHERE estado IN ('pendiente','en_proceso')
          AND fecha_programada >= date('now')
          AND fecha_programada <= date(?)
    """, (fecha_limite,)).fetchall()
    productos_en_plan_set = {p[0] for p in productos_en_plan_bd if p[0]}

    # Sumar Google Calendar
    eventos_cal = _calendar_events_cached()
    fecha_hoy = datetime.now().date()
    fecha_top = fecha_hoy + timedelta(days=dias)
    for ev in eventos_cal:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            if f < fecha_hoy or f > fecha_top:
                continue
        except Exception:
            continue
        titulo = (ev.get('titulo') or '').upper()
        # Match contra productos
        for prod_upper in productos_total_set:
            palabras = [w for w in prod_upper.split() if len(w) > 4]
            if prod_upper in titulo or any(w in titulo for w in palabras):
                productos_en_plan_set.add(prod_upper)

    productos_con_config = c.execute("""
        SELECT UPPER(TRIM(producto_nombre)) FROM sku_planeacion_config WHERE activo=1
    """).fetchall()
    productos_config_set = {p[0] for p in productos_con_config if p[0]}

    # Productos que están en formula pero NO en plan
    sin_plan = []
    for p in productos_total:
        if p[0] and (p[0].strip().upper() not in productos_en_plan_set):
            sin_plan.append(p[0])

    return jsonify({
        'horizonte_dias': dias,
        'total_skus': len(productos_total_set),
        'en_plan': len(productos_en_plan_set),
        'en_plan_futuro': len(productos_en_plan_set),  # back-compat
        'con_config': len(productos_config_set),
        'sin_plan': sorted(sin_plan),
        'cobertura_pct': round(len(productos_en_plan_set) / max(1, len(productos_total_set)) * 100, 1),
    })


@bp.route('/api/planta/producto-nuevo', methods=['POST'])
def producto_nuevo_rapido():
    """Sebastian (30-abr-2026): "debe permitir adicionar nuevos productos
    porque se vienen lanzamientos... lo palancamos desde área científica
    pero debe existir en producción también por si se necesita programar
    algo de manera prioritaria".

    Crea de un solo click:
      1. Entrada en formula_headers (sin items)
      2. sku_planeacion_config con cadencia=null (auto por umbral)
      3. Opcionalmente programa una producción inicial prioritaria

    Body: {producto_nombre*, lote_size_kg*, categoria?, cadencia_dias?,
           merma_pct?, fecha_primera_prod?, lotes_inicial?,
           prioritario? (boolean)}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    producto = (d.get('producto_nombre') or '').strip()
    lote_kg = d.get('lote_size_kg')
    if not producto or not lote_kg:
        return jsonify({'error': 'producto_nombre y lote_size_kg requeridos'}), 400
    lote_kg = float(lote_kg)
    conn = get_db(); c = conn.cursor()

    # 1. formula_headers
    existing_fh = c.execute(
        "SELECT producto_nombre FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    if not existing_fh:
        c.execute("""
            INSERT INTO formula_headers (producto_nombre, lote_size_kg, unidad_base_g)
            VALUES (?, ?, 1.0)
        """, (producto, lote_kg))
        log_msg_fh = 'fórmula creada'
    else:
        log_msg_fh = 'fórmula ya existía'

    # 2. sku_planeacion_config
    existing_cfg = c.execute(
        "SELECT id FROM sku_planeacion_config WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    if not existing_cfg:
        c.execute("""
            INSERT INTO sku_planeacion_config
              (producto_nombre, categoria, cadencia_dias, cobertura_target_dias,
               cobertura_min_dias, cobertura_max_dias, merma_pct, prioridad,
               activo, notas)
            VALUES (?, ?, ?, 60, 20, 90, ?, 2, 1, ?)
        """, (
            producto, d.get('categoria'), d.get('cadencia_dias'),
            float(d.get('merma_pct') or 5.0),
            f'Creado por {user} el {datetime.now().date()} (lanzamiento)',
        ))
        log_msg_cfg = 'config creada'
    else:
        log_msg_cfg = 'config ya existía'

    # 3. Producción inicial prioritaria (opcional)
    produccion_id = None
    if d.get('prioritario') or d.get('fecha_primera_prod'):
        from datetime import datetime as _dt2
        fecha_obj = (d.get('fecha_primera_prod') or '').strip()
        if not fecha_obj:
            # Próximo L/M/V
            fobj = _dt2.now().date() + timedelta(days=2)
            while fobj.weekday() not in (0, 2, 4):
                fobj += timedelta(days=1)
            fecha_obj = fobj.isoformat()
        merma = float(d.get('merma_pct') or 5.0)
        kg_con_merma = lote_kg * (1 + merma / 100.0)
        cur = c.execute("""
            INSERT INTO produccion_programada
              (producto, fecha_programada, lotes, estado, observaciones, origen, cantidad_kg)
            VALUES (?, ?, ?, 'pendiente', ?, 'producto_nuevo_lanzamiento', ?)
        """, (
            producto, fecha_obj, int(d.get('lotes_inicial') or 1),
            f'PRODUCTO NUEVO · creado por {user} · {("prioritario" if d.get("prioritario") else "primera producción")}',
            kg_con_merma,
        ))
        produccion_id = cur.lastrowid
        # Auto-asignar área + operarios al producto nuevo
        try:
            cap_min = kg_con_merma * 1.2
            tanques = c.execute("""
                SELECT area_codigo FROM equipos_planta
                WHERE activo=1 AND tipo IN ('tanque','marmita','olla')
                  AND capacidad_litros >= ?
                ORDER BY capacidad_litros ASC LIMIT 1
            """, (cap_min,)).fetchone()
            if tanques:
                area_chk = c.execute(
                    "SELECT id FROM areas_planta WHERE codigo=? AND puede_producir=1 AND activo=1",
                    (tanques[0],)
                ).fetchone()
                if area_chk:
                    c.execute("UPDATE produccion_programada SET area_id=? WHERE id=?",
                              (area_chk[0], produccion_id))
            _asignar_operarios_a_produccion(conn, produccion_id)
        except Exception:
            pass

    # Notificar a admins
    try:
        from blueprints.notif import push_notif_multi
        push_notif_multi(
            ['sebastian', 'alejandro'], 'planta',
            f'🆕 Producto nuevo: {producto}',
            body=f'Lote {lote_kg}kg · creado por {user}' + (f' · 1ra producción {fecha_obj}' if produccion_id else ''),
            link='/inventarios#programacion',
            remitente=user,
        )
    except Exception:
        pass

    conn.commit()
    return jsonify({
        'ok': True,
        'producto': producto,
        'formula': log_msg_fh,
        'config': log_msg_cfg,
        'produccion_creada_id': produccion_id,
        'mensaje': f'✓ {producto} agregado al sistema',
    })


@bp.route('/api/auto-plan/aprender-historico', methods=['GET'])
def aprender_historico():
    """Analiza producciones histórias (BD + Google Calendar) y deriva
    cadencias reales. Devuelve comparación con la config actual.

    Querystring: ?meses_atras=12 (default — cuánto histórico considerar)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        meses_atras = int(request.args.get('meses_atras', 12))
    except Exception:
        meses_atras = 12

    conn = get_db(); c = conn.cursor()
    fecha_desde = (datetime.now() - timedelta(days=meses_atras * 30)).date()

    # 1. Producciones desde produccion_programada (BD interna)
    rows_bd = c.execute("""
        SELECT producto, fecha_programada, lotes, estado, cantidad_kg
        FROM produccion_programada
        WHERE date(fecha_programada) >= ?
          AND estado IN ('completado','en_proceso','pendiente')
          AND producto IS NOT NULL
        ORDER BY producto, fecha_programada
    """, (fecha_desde.isoformat(),)).fetchall()

    # Acumular fechas por producto
    fechas_por_producto = {}
    for prod, fecha, lotes, estado, kg in rows_bd:
        try:
            f = datetime.strptime(fecha[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        key = (prod or '').strip().upper()
        if not key:
            continue
        fechas_por_producto.setdefault(key, {'nombre_real': prod, 'fechas': [],
                                              'lotes_total': 0, 'kg_total': 0})
        fechas_por_producto[key]['fechas'].append(f)
        fechas_por_producto[key]['lotes_total'] += (lotes or 1)
        fechas_por_producto[key]['kg_total'] += (kg or 0)

    # 2. Producciones desde Google Calendar (eventos con producto en el título)
    try:
        from blueprints.programacion import _fetch_calendar_events
        cal_events = _fetch_calendar_events(days_ahead=0) or []
        # _fetch_calendar_events devuelve eventos futuros, pero podemos
        # ampliar para incluir el pasado. Si no, usamos solo BD.
    except Exception:
        cal_events = []

    # Cargar productos en formula_headers para detectar nuevos
    productos_formula = c.execute("""
        SELECT producto_nombre, lote_size_kg FROM formula_headers
    """).fetchall()
    productos_formula_set = {(p[0] or '').strip().upper(): {'nombre': p[0], 'lote_kg': p[1]}
                             for p in productos_formula if p[0]}

    # 3. Cargar config actual
    configs = c.execute("""
        SELECT producto_nombre, cadencia_dias, cobertura_target_dias, merma_pct
        FROM sku_planeacion_config WHERE activo=1
    """).fetchall()
    config_por_producto = {(c[0] or '').strip().upper(): {
        'cadencia': c[1], 'target': c[2], 'merma': c[3]
    } for c in configs}

    # 4. Análisis por producto
    aprendizaje = []
    productos_con_historico = set()
    for key, info in fechas_por_producto.items():
        productos_con_historico.add(key)
        cadencia_real = _calcular_cadencia_real(info['fechas'])
        cfg = config_por_producto.get(key, {})
        cadencia_config = cfg.get('cadencia')
        ultima = max(info['fechas']) if info['fechas'] else None
        primera = min(info['fechas']) if info['fechas'] else None
        # Diferencia entre real vs config
        diferencia = None
        if cadencia_real and cadencia_config:
            diferencia = cadencia_real - cadencia_config
        # Recomendación
        recomendar_cambiar = False
        if cadencia_real and (not cadencia_config or abs((cadencia_real or 0) - (cadencia_config or 0)) > 7):
            recomendar_cambiar = True
        aprendizaje.append({
            'producto': info['nombre_real'],
            'lotes_historicos': len(info['fechas']),
            'lotes_total_unidades': info['lotes_total'],
            'kg_total': round(info['kg_total'], 1),
            'primera_produccion': primera.isoformat() if primera else None,
            'ultima_produccion': ultima.isoformat() if ultima else None,
            'dias_desde_ultima': (datetime.now().date() - ultima).days if ultima else None,
            'cadencia_real_dias': cadencia_real,
            'cadencia_configurada': cadencia_config,
            'diferencia_dias': diferencia,
            'recomendar_actualizar': recomendar_cambiar,
            'tiene_config': key in config_por_producto,
            'tiene_formula': key in productos_formula_set,
        })

    # 5. Productos NUEVOS (en formula_headers pero sin histórico)
    productos_nuevos = []
    for key, info in productos_formula_set.items():
        if key not in productos_con_historico:
            cfg = config_por_producto.get(key, {})
            productos_nuevos.append({
                'producto': info['nombre'],
                'lote_kg': info['lote_kg'],
                'tiene_config': key in config_por_producto,
                'cadencia_configurada': cfg.get('cadencia'),
                'sugerencia': 'Producir primer lote pronto — sin histórico',
            })

    return jsonify({
        'fecha_analisis': datetime.now().isoformat(),
        'meses_atras': meses_atras,
        'aprendizaje': sorted(aprendizaje, key=lambda x: -(x.get('lotes_historicos') or 0)),
        'productos_nuevos': productos_nuevos,
        'kpis': {
            'productos_con_historico': len(aprendizaje),
            'productos_nuevos_sin_historico': len(productos_nuevos),
            'recomendaciones_actualizar': sum(1 for a in aprendizaje if a['recomendar_actualizar']),
            'total_lotes_analizados': sum(a['lotes_historicos'] for a in aprendizaje),
        },
    })


@bp.route('/api/auto-plan/aplicar-aprendizaje', methods=['POST'])
def aplicar_aprendizaje():
    """Toma las cadencias detectadas del histórico y actualiza la config.
    Body: {productos: [{producto, cadencia_real_dias}, ...]}
    O sin body para aplicar TODAS las recomendaciones.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    productos_a_aplicar = d.get('productos')

    conn = get_db(); c = conn.cursor()

    if not productos_a_aplicar:
        # Modo sin lista: tomar todas las recomendaciones del análisis
        from flask import url_for
        analisis = aprender_historico()
        if hasattr(analisis, 'get_json'):
            data = analisis.get_json()
        else:
            data = analisis[0].get_json() if isinstance(analisis, tuple) else {}
        productos_a_aplicar = []
        for a in (data.get('aprendizaje') or []):
            if a.get('recomendar_actualizar') and a.get('cadencia_real_dias'):
                productos_a_aplicar.append({
                    'producto': a['producto'],
                    'cadencia_real_dias': a['cadencia_real_dias'],
                })

    actualizados = []
    for it in productos_a_aplicar:
        producto = (it.get('producto') or '').strip()
        cadencia = it.get('cadencia_real_dias')
        if not producto or not cadencia:
            continue
        # Verificar si tiene config
        existing = c.execute(
            "SELECT id, cadencia_dias FROM sku_planeacion_config WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND activo=1",
            (producto,)
        ).fetchone()
        if existing:
            c.execute("""
                UPDATE sku_planeacion_config SET
                  cadencia_dias=?, actualizado_en=datetime('now'),
                  notas=COALESCE(notas,'')||' · Aprendido del histórico '||date('now')
                WHERE id=?
            """, (int(cadencia), existing[0]))
            actualizados.append({'producto': producto, 'cadencia_anterior': existing[1],
                                  'cadencia_nueva': cadencia, 'accion': 'actualizado'})
        else:
            # Crear config nueva con cadencia detectada
            c.execute("""
                INSERT INTO sku_planeacion_config
                  (producto_nombre, cadencia_dias, cobertura_target_dias,
                   cobertura_min_dias, merma_pct, prioridad, activo, notas)
                VALUES (?, ?, 60, 20, 5.0, 3, 1, ?)
            """, (producto, int(cadencia),
                  f'Auto-creado desde histórico {datetime.now().date()}'))
            actualizados.append({'producto': producto, 'cadencia_anterior': None,
                                  'cadencia_nueva': cadencia, 'accion': 'creado'})

    conn.commit()
    return jsonify({
        'ok': True, 'actualizados': actualizados, 'total': len(actualizados),
    })


# ════════════════════════════════════════════════════════════════════════
# MAQUILA INTELIGENTE — clientes + pedidos integrados al motor del Plan
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "Kelly Guerra compra productos para marca de ella
# pero misma fórmula Animus... espacio de maquila inteligente, si Fernando
# lleva 500 le adiciona a la producción esas 500 unidades".

@bp.route('/api/maquila/clientes', methods=['GET', 'POST'])
def maquila_clientes():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        nombre = (d.get('nombre') or '').strip()
        if not nombre:
            return jsonify({'error': 'nombre requerido'}), 400
        try:
            c.execute("""
                INSERT INTO clientes_maquila
                  (nombre, nit_cedula, email, telefono, es_marca_propia,
                   empresa_grupo, comparte_formula_con, margen_seguridad_pct, notas)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                nombre, d.get('nit_cedula'), d.get('email'), d.get('telefono'),
                1 if d.get('es_marca_propia') else 0,
                d.get('empresa_grupo'), d.get('comparte_formula_con'),
                int(d.get('margen_seguridad_pct') or 5),
                d.get('notas'),
            ))
            conn.commit()
            return jsonify({'ok': True, 'id': c.lastrowid})
        except sqlite3.IntegrityError:
            return jsonify({'error': 'Cliente ya existe con ese nombre'}), 409
    rows = c.execute(
        "SELECT * FROM clientes_maquila WHERE activo=1 ORDER BY nombre"
    ).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'clientes': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/maquila/pedidos', methods=['GET', 'POST'])
def maquila_pedidos():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        cliente_id = d.get('cliente_id')
        producto = (d.get('producto_nombre') or '').strip()
        unidades = int(d.get('unidades') or 0)
        if not cliente_id or not producto or unidades <= 0:
            return jsonify({'error': 'cliente_id, producto_nombre, unidades requeridos'}), 400
        # Buscar cliente
        cli = c.execute(
            "SELECT nombre FROM clientes_maquila WHERE id=?", (cliente_id,)
        ).fetchone()
        if not cli:
            return jsonify({'error': 'Cliente no existe'}), 400
        # Generar número
        n = c.execute(
            "SELECT COALESCE(MAX(CAST(SUBSTR(numero,4) AS INTEGER)),0)+1 FROM maquila_pedidos WHERE numero LIKE 'MQ-%'"
        ).fetchone()[0] or 1
        numero = f'MQ-{n:04d}'
        # Calcular kg estimados desde presentación
        kg_est = d.get('kg_estimados')
        if not kg_est and d.get('presentacion_id'):
            pr = c.execute(
                "SELECT factor_g_por_unidad FROM producto_presentaciones WHERE id=?",
                (d.get('presentacion_id'),)
            ).fetchone()
            if pr and pr[0]:
                kg_est = (unidades * pr[0]) / 1000.0
        try:
            c.execute("""
                INSERT INTO maquila_pedidos
                  (numero, cliente_id, cliente_nombre, producto_nombre, presentacion_id,
                   unidades, kg_estimados, fecha_entrega_objetivo, precio_unidad,
                   observaciones, creado_por)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                numero, cliente_id, cli[0], producto,
                d.get('presentacion_id'),
                unidades, kg_est,
                (d.get('fecha_entrega_objetivo') or '').strip() or None,
                d.get('precio_unidad'),
                d.get('observaciones'),
                user,
            ))
            conn.commit()
            return jsonify({'ok': True, 'numero': numero, 'id': c.lastrowid,
                            'kg_estimados': kg_est})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    # GET — filtros opcionales
    estado = (request.args.get('estado') or '').strip()
    cliente = request.args.get('cliente_id')
    where = []
    params = []
    if estado and estado != 'todos':
        where.append('mp.estado=?'); params.append(estado)
    if cliente:
        where.append('mp.cliente_id=?'); params.append(int(cliente))
    sql = """
        SELECT mp.*, cm.nombre as cliente_nombre_full,
               pp.fecha_programada as produccion_fecha,
               pp.estado as produccion_estado
        FROM maquila_pedidos mp
        LEFT JOIN clientes_maquila cm ON cm.id = mp.cliente_id
        LEFT JOIN produccion_programada pp ON pp.id = mp.produccion_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY mp.fecha_entrega_objetivo ASC NULLS LAST, mp.id DESC"
    rows = c.execute(sql, params).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    # KPIs
    pendientes = sum(1 for it in items if it['estado'] in ('recibido', 'planificado'))
    en_prod = sum(1 for it in items if it['estado'] == 'en_produccion')
    return jsonify({
        'pedidos': items, 'total': len(items),
        'pendientes': pendientes, 'en_produccion': en_prod,
    })


@bp.route('/api/maquila/pedidos/<int:pedido_id>/asignar-produccion', methods=['POST'])
def maquila_asignar_produccion(pedido_id):
    """Asocia un pedido de maquila a una producción específica."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    prod_id = d.get('produccion_id')
    if not prod_id:
        return jsonify({'error': 'produccion_id requerido'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("""
        UPDATE maquila_pedidos SET produccion_id=?, estado='planificado', actualizado_en=datetime('now')
        WHERE id=?
    """, (prod_id, pedido_id))
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/maquila/pedidos/<int:pedido_id>', methods=['DELETE'])
def maquila_pedido_cancelar(pedido_id):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute(
        "UPDATE maquila_pedidos SET estado='cancelado' WHERE id=?", (pedido_id,)
    )
    conn.commit()
    return jsonify({'ok': True})


# ════════════════════════════════════════════════════════════════════════
# Asignación automática de operarios (Mayerlin fija + rotación inteligente)
# ════════════════════════════════════════════════════════════════════════
def _asignar_operarios_a_produccion(conn, produccion_id):
    """Asigna automáticamente operarios a una producción según reglas:
       - Dispensación: SIEMPRE Mayerlin (regla dura)
       - Elaboración/Envasado/Acondicionamiento: rotan según historial reciente
    """
    c = conn.cursor()
    # Buscar Mayerlin
    mayerlin = c.execute(
        "SELECT id FROM operarios_planta WHERE LOWER(nombre)='mayerlin' AND activo=1 LIMIT 1"
    ).fetchone()
    op_disp_id = mayerlin[0] if mayerlin else None

    # Para las otras fases: encontrar operario con MENOS asignaciones en últimos 7 días
    def operario_menos_asignado(rol_excluir=None):
        rows = c.execute(f"""
            SELECT op.id, op.nombre,
                   (SELECT COUNT(*) FROM produccion_programada pp
                    WHERE pp.fecha_programada >= date('now','-7 days')
                      AND (pp.operario_elaboracion_id=op.id
                        OR pp.operario_envasado_id=op.id
                        OR pp.operario_acondicionamiento_id=op.id)) AS carga
            FROM operarios_planta op
            WHERE op.activo=1
              AND op.fija_en_dispensacion=0
              AND op.es_jefe_produccion=0
              AND LOWER(op.nombre) != 'mayerlin'
            ORDER BY carga ASC, op.nombre LIMIT 1
        """).fetchone()
        return rows[0] if rows else None

    op_elab_id = operario_menos_asignado()
    op_env_id = operario_menos_asignado()
    op_acond_id = operario_menos_asignado()

    c.execute("""
        UPDATE produccion_programada SET
          operario_dispensacion_id = COALESCE(operario_dispensacion_id, ?),
          operario_elaboracion_id = COALESCE(operario_elaboracion_id, ?),
          operario_envasado_id = COALESCE(operario_envasado_id, ?),
          operario_acondicionamiento_id = COALESCE(operario_acondicionamiento_id, ?)
        WHERE id=?
    """, (op_disp_id, op_elab_id, op_env_id, op_acond_id, produccion_id))


@bp.route('/api/planta/produccion/<int:prod_id>/asignar-operarios-auto', methods=['POST'])
def asignar_operarios_auto(prod_id):
    """Asigna operarios automáticamente (Mayerlin fija + rotación)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    _asignar_operarios_a_produccion(conn, prod_id)
    conn.commit()
    return jsonify({'ok': True, 'mensaje': 'Operarios asignados'})


@bp.route('/api/planta/asignar-operarios-bulk', methods=['POST'])
def asignar_operarios_bulk():
    """Asigna operarios automáticamente a TODAS las producciones pendientes
    sin operarios. Útil después de Auto-Plan."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT id FROM produccion_programada
        WHERE estado IN ('pendiente','en_proceso')
          AND fecha_programada >= date('now')
          AND (operario_dispensacion_id IS NULL
            OR operario_elaboracion_id IS NULL
            OR operario_envasado_id IS NULL)
    """).fetchall()
    asignadas = 0
    for (pid,) in rows:
        _asignar_operarios_a_produccion(conn, pid)
        asignadas += 1
    conn.commit()
    return jsonify({'ok': True, 'asignadas': asignadas})


# ════════════════════════════════════════════════════════════════════════
# Descarga Plan PDF/Excel (para Alejandro)
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/plan/exportar', methods=['GET'])
def plan_exportar():
    """Exporta el plan a Excel para revisión.
    Querystring: ?meses=1|2|3|6|12 ?formato=xlsx|csv (default xlsx)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from flask import Response
    try:
        meses = int(request.args.get('meses', 3))
    except Exception:
        meses = 3
    formato = (request.args.get('formato') or 'xlsx').lower()

    conn = get_db(); c = conn.cursor()
    proyeccion = _generar_proyeccion_lotes(c, meses)

    if formato == 'csv':
        import io
        buffer = io.StringIO()
        buffer.write('Producto,Fecha,Mes,Lote_kg,Kg_con_merma,Velocidad_dia,Tipo,Cadencia_dias\n')
        for p in proyeccion:
            buffer.write(f"{p['producto']},{p['fecha']},{p['mes']},{p['lote_kg']},{p['kg_con_merma']:.1f},{p['velocidad_dia']:.2f},{p['tipo']},{p.get('cadencia_dias') or ''}\n")
        return Response(
            buffer.getvalue(), mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=plan_{meses}m.csv'}
        )

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl no disponible'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = f'Plan {meses}m'

    # Header
    headers = ['Producto', 'Fecha', 'Mes', 'Lote kg', 'Kg con merma',
               'Velocidad/día', 'Tipo', 'Cadencia']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for r, p in enumerate(proyeccion, 2):
        ws.cell(row=r, column=1, value=p['producto'])
        ws.cell(row=r, column=2, value=p['fecha'])
        ws.cell(row=r, column=3, value=p['mes'])
        ws.cell(row=r, column=4, value=p['lote_kg'])
        ws.cell(row=r, column=5, value=round(p['kg_con_merma'], 1))
        ws.cell(row=r, column=6, value=round(p['velocidad_dia'], 2))
        ws.cell(row=r, column=7, value=p['tipo'])
        ws.cell(row=r, column=8, value=p.get('cadencia_dias') or '')

    # Anchos
    for col, w in zip('ABCDEFGH', [42, 12, 10, 10, 14, 14, 14, 12]):
        ws.column_dimensions[col].width = w

    # Hoja resumen mensual
    ws2 = wb.create_sheet('Resumen mensual')
    ws2.cell(row=1, column=1, value='Mes').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Lotes').font = Font(bold=True)
    ws2.cell(row=1, column=3, value='Kg total').font = Font(bold=True)
    ws2.cell(row=1, column=4, value='SKUs distintos').font = Font(bold=True)
    resumen = {}
    for p in proyeccion:
        m = p['mes']
        if m not in resumen:
            resumen[m] = {'lotes': 0, 'kg': 0, 'skus': set()}
        resumen[m]['lotes'] += 1
        resumen[m]['kg'] += p['kg_con_merma']
        resumen[m]['skus'].add(p['producto'])
    for r, m in enumerate(sorted(resumen), 2):
        ws2.cell(row=r, column=1, value=m)
        ws2.cell(row=r, column=2, value=resumen[m]['lotes'])
        ws2.cell(row=r, column=3, value=round(resumen[m]['kg'], 1))
        ws2.cell(row=r, column=4, value=len(resumen[m]['skus']))

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=plan_{meses}m_alejandro.xlsx'}
    )


@bp.route('/api/planta/produccion/<int:prod_id>/aceptar-recomendacion', methods=['POST'])
def aceptar_recomendacion(prod_id):
    """Acepta una recomendación de mover producción a otra fecha."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    nueva_fecha = (d.get('nueva_fecha') or '').strip()
    if not nueva_fecha:
        return jsonify({'error': 'nueva_fecha requerida'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("""
        UPDATE produccion_programada SET
          fecha_programada=?,
          observaciones=COALESCE(observaciones,'')||' | Movida por '||?||' a '||?||' (recomendación demanda)'
        WHERE id=?
    """, (nueva_fecha, user, nueva_fecha, prod_id))
    conn.commit()
    return jsonify({'ok': True, 'produccion_id': prod_id, 'nueva_fecha': nueva_fecha})


@bp.route('/api/auto-plan/runs', methods=['GET'])
def auto_plan_runs():
    """Histórico de ejecuciones del auto-plan."""
    u, err, code = _auth()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT id, ejecutado_at, ejecutado_por, tipo, horizonte_dias,
               producciones_creadas, compras_creadas, alertas_criticas,
               emails_enviados, error, duracion_ms
        FROM auto_plan_runs
        ORDER BY ejecutado_at DESC
        LIMIT 30
    """).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'runs': [dict(zip(cols, r)) for r in rows]})


# ── Configs CRUD ──────────────────────────────────────────────────────

@bp.route('/api/auto-plan/configs/sku', methods=['GET'])
def configs_sku():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT spc.*, fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
        ORDER BY spc.prioridad, spc.producto_nombre
    """).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'configs': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/auto-plan/configs/sku/<int:config_id>', methods=['PUT'])
def configs_sku_update(config_id):
    u, err, code = _auth()
    if err:
        return err, code
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    campos = []
    params = []
    for col in ('cadencia_dias', 'cobertura_target_dias', 'cobertura_min_dias',
                'cobertura_max_dias', 'merma_pct', 'prioridad', 'categoria',
                'presentacion_default_id', 'notas', 'alias_calendar'):
        if col in d:
            campos.append(f'{col}=?')
            params.append(d[col])
    if not campos:
        return jsonify({'error': 'Sin cambios'}), 400
    campos.append("actualizado_en=datetime('now')")
    params.append(config_id)
    c.execute(f"UPDATE sku_planeacion_config SET {', '.join(campos)} WHERE id=?", params)
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/auto-plan/configs/mp', methods=['GET', 'POST'])
def configs_mp():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        c.execute("""
            INSERT OR REPLACE INTO mp_lead_time_config
              (material_id, material_nombre, proveedor_principal, lead_time_dias,
               buffer_dias, cobertura_min_dias, cobertura_ideal_dias, origen,
               es_envase, activo, actualizado_en)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, datetime('now'))
        """, (
            d.get('material_id'), d.get('material_nombre'),
            d.get('proveedor_principal'), int(d.get('lead_time_dias') or 14),
            int(d.get('buffer_dias') or 30), int(d.get('cobertura_min_dias') or 30),
            int(d.get('cobertura_ideal_dias') or 60),
            d.get('origen') or 'local', 1 if d.get('es_envase') else 0,
        ))
        conn.commit()
        return jsonify({'ok': True})
    rows = c.execute("SELECT * FROM mp_lead_time_config WHERE activo=1 ORDER BY origen, material_nombre").fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'configs': [dict(zip(cols, r)) for r in rows]})


# ════════════════════════════════════════════════════════════════════════
# Asistente conversacional · Claude API con contexto de planta
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "asistente conversacional ya tenemos la cosa
# flotante" — conectar el chat flotante con Claude API + contexto de
# planta para que pueda responder preguntas como:
#   "¿Cuánto Suero AH puedo producir esta semana?"
#   "¿Por qué hay alerta crítica?"
#   "Programa 2 lotes de Vit C para el viernes"

def _build_planta_context(c):
    """Recolecta snapshot del estado de planta para el prompt de Claude."""
    snapshot = {}

    # KPIs
    try:
        snapshot['areas_planta'] = [
            dict(zip(['codigo', 'nombre', 'estado', 'puede_producir', 'puede_envasar'], r))
            for r in c.execute(
                "SELECT codigo, nombre, estado, puede_producir, puede_envasar "
                "FROM areas_planta WHERE activo=1 ORDER BY orden LIMIT 15"
            ).fetchall()
        ]
    except Exception:
        snapshot['areas_planta'] = []

    try:
        snapshot['equipos_resumen'] = [
            {'area_codigo': r[0], 'tipo': r[1], 'count': r[2]}
            for r in c.execute(
                "SELECT area_codigo, tipo, COUNT(*) FROM equipos_planta "
                "WHERE activo=1 AND tipo IN ('tanque','marmita','olla','envasadora') "
                "GROUP BY area_codigo, tipo"
            ).fetchall()
        ]
    except Exception:
        snapshot['equipos_resumen'] = []

    try:
        snapshot['operarios'] = [
            dict(zip(['nombre', 'rol'], r))
            for r in c.execute(
                "SELECT nombre, rol_predeterminado FROM operarios_planta WHERE activo=1"
            ).fetchall()
        ]
    except Exception:
        snapshot['operarios'] = []

    # Producciones próximas 14d
    try:
        snapshot['producciones_proximas'] = [
            {'producto': r[0], 'fecha': r[1], 'lotes': r[2], 'estado': r[3]}
            for r in c.execute(
                "SELECT producto, fecha_programada, lotes, estado FROM produccion_programada "
                "WHERE fecha_programada >= date('now') AND fecha_programada <= date('now','+14 days') "
                "ORDER BY fecha_programada ASC LIMIT 25"
            ).fetchall()
        ]
    except Exception:
        snapshot['producciones_proximas'] = []

    # Cadencias configuradas
    try:
        snapshot['cadencias_skus'] = [
            {'producto': r[0], 'cadencia_d': r[1], 'cobertura_d': r[2], 'merma_pct': r[3]}
            for r in c.execute(
                "SELECT producto_nombre, cadencia_dias, cobertura_target_dias, merma_pct "
                "FROM sku_planeacion_config WHERE activo=1 AND cadencia_dias IS NOT NULL "
                "ORDER BY prioridad LIMIT 10"
            ).fetchall()
        ]
    except Exception:
        snapshot['cadencias_skus'] = []

    # Último auto-plan run
    try:
        last_run = c.execute(
            "SELECT ejecutado_at, producciones_creadas, compras_creadas, alertas_criticas "
            "FROM auto_plan_runs ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if last_run:
            snapshot['ultimo_run'] = {
                'fecha': last_run[0], 'prods': last_run[1],
                'compras': last_run[2], 'alertas': last_run[3]
            }
    except Exception:
        pass

    return snapshot


@bp.route('/api/asistente/planta', methods=['POST'])
def asistente_planta():
    """Asistente conversacional con contexto de planta. Usa Claude API.
    Body: {pregunta: str, historial?: [{role, content}, ...]}"""
    import os
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    pregunta = (d.get('pregunta') or '').strip()
    if not pregunta:
        return jsonify({'error': 'pregunta requerida'}), 400
    historial = d.get('historial') or []

    api_key = os.environ.get('ANTHROPIC_API_KEY') or os.environ.get('CLAUDE_API_KEY')
    if not api_key:
        return jsonify({
            'respuesta': 'El asistente conversacional necesita ANTHROPIC_API_KEY configurada en Render. Cuando esté lista, podré responder con contexto completo de la planta.',
            'sin_api_key': True,
        })

    conn = get_db(); c = conn.cursor()
    try:
        ctx = _build_planta_context(c)
    except Exception as e:
        ctx = {}
        log.warning(f'asistente: no se pudo armar contexto: {e}')

    es_admin = user in ADMIN_USERS
    system_prompt = f"""Eres EOS Planta, el asistente experto de planta para Espagiria Laboratorios (HHA Group), laboratorio cosmético colombiano certificado INVIMA. Te creó Claude. Eres preciso, breve y útil — operativo, no chatbot genérico.

CONTEXTO ACTUAL (snapshot {datetime.now().isoformat()}):
{json.dumps(ctx, ensure_ascii=False, indent=2)}

REGLAS DE PLANEACIÓN:
- Vit C → cadencia 30d (oxida)
- Suero AH 1.5% → cadencia 90d, lote 90kg
- L/M/V producir · Ma/Ju acondicionar/envasar/conteo cíclico
- MP mínimo 30d, ideal 60d · Envases mínimo 90d (China lead 180d)
- 4 operarios: Mayerlin (dispensación fija), Camilo, Milton, Sebastián M.
- Áreas: FAB1/2/3, ENV1/2, DISP, LAV, ESC1, ACOND, FAB_FLOAT, CC, RECEP

ACCIONES DISPONIBLES (puedes sugerir al usuario que las dispare):
- Ejecutar Auto-Plan ahora → /api/auto-plan/aplicar (solo admin)
- Ver alertas críticas activas → tab "Plan Semanal" o /api/asistente/tool/listar_alertas
- Ver producciones próximas → tab "Plan Semanal" o /api/asistente/tool/listar_producciones_proximas
- Configurar cadencia → tab "🤖 Auto-Plan" → "Cadencias por SKU"
- Activar/desactivar cron → tab "🤖 Auto-Plan" → toggle (solo admin)

Usuario actual: {user} ({'ADMIN' if es_admin else 'usuario'})

INSTRUCCIONES:
- Responde en español, conciso (máximo 200 palabras salvo que pidan detalle).
- Si la pregunta es operativa (cuánto, cuándo, dónde), USA el contexto del snapshot para contestar con datos reales.
- Si te piden hacer algo destructivo (ejecutar, crear), describe los pasos y dile dónde hacer click — NO inventes acciones.
- Si no hay datos suficientes, dilo: "no tengo info sobre X — revisa Y".
- Si detectas riesgo (alertas críticas, MP en déficit, sala sin limpieza), señálalo proactivamente.
"""

    messages = []
    for h in historial[-10:]:
        if h.get('role') in ('user', 'assistant') and h.get('content'):
            messages.append({'role': h['role'], 'content': h['content']})
    messages.append({'role': 'user', 'content': pregunta})

    try:
        import urllib.request, urllib.error
        body = json.dumps({
            'model': 'claude-haiku-4-5-20251001',
            'max_tokens': 800,
            'system': system_prompt,
            'messages': messages,
        }).encode('utf-8')
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=body, method='POST',
            headers={
                'content-type': 'application/json',
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
            },
        )
        with urllib.request.urlopen(req, timeout=45) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        text = (data.get('content') or [{}])[0].get('text', '').strip()
        return jsonify({
            'respuesta': text,
            'usage': data.get('usage', {}),
            'modelo': 'claude-haiku-4-5',
        })
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode('utf-8')
        except Exception:
            err_body = str(e)
        log.warning(f'Claude API error: {err_body}')
        return jsonify({'error': f'Claude API: {e.code}', 'detalle': err_body[:200]}), 500
    except Exception as e:
        log.warning(f'Asistente fallo: {e}')
        return jsonify({'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# Cron control desde UI (sin env var)
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/auto-plan/cron/state', methods=['GET'])
def cron_state():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    r = c.execute("SELECT habilitado, activado_por, activado_at, ultima_ejecucion_at, errores_consecutivos, notas FROM auto_plan_cron_state WHERE id=1").fetchone()
    if not r:
        return jsonify({'habilitado': False})
    return jsonify({
        'habilitado': bool(r[0]),
        'activado_por': r[1],
        'activado_at': r[2],
        'ultima_ejecucion_at': r[3],
        'errores_consecutivos': r[4],
        'notas': r[5],
    })


@bp.route('/api/auto-plan/cron/toggle', methods=['POST'])
def cron_toggle():
    u, err, code = _auth()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    d = request.json or {}
    habilitar = bool(d.get('habilitar'))
    conn = get_db(); c = conn.cursor()
    if habilitar:
        c.execute("""
            UPDATE auto_plan_cron_state SET
              habilitado=1, activado_por=?, activado_at=datetime('now'),
              errores_consecutivos=0, notas=?
            WHERE id=1
        """, (u, 'Activado desde UI'))
    else:
        c.execute("UPDATE auto_plan_cron_state SET habilitado=0, notas=? WHERE id=1",
                  ('Desactivado desde UI',))
    conn.commit()
    return jsonify({'ok': True, 'habilitado': habilitar})


# ════════════════════════════════════════════════════════════════════════
# FORECAST BLACK FRIDAY — pre-stock necesario por SKU para nov-dic
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "Black Friday 2025 fue catástrofe por escasez.
# Para BF 2026 hay que pre-stockear MUCHO más". Multiplicadores REALES
# calculados de venta nov-2025 vs venta normal 2026.

# Multiplicadores empíricos durante 14 días pico BF 2025
BF_MULTIPLIERS = {
    'GLOSSN':       6.6,
    'CMULP':        5.1,
    'MAXLASH':      4.9,
    'HKJ':          4.8,
    'GELH':         4.5,
    'SMULPP':       4.0,
    'CCAFE':        3.9,
    'CRB3BHA':      3.9,
    'NPHA10':       3.4,
    'TRX':          2.8,
    'TRX10':        2.8,
    'LAH':          2.6,
    'LKJ':          2.5,
    'CRETT':        2.5,
    'LBHA':         2.5,
    'RECN-2':       2.4,
    'ECENT':        2.2,
    'EMLIM':        2.1,
    'NIA':          2.0,
    'NIA10':        2.0,
    'SVITC33':      2.0,
    'BHA33':        1.9,
    'SAH10':        1.8,
    'AZHC30':       1.8,
    'SAH':          1.8,
    'TRIAC':        1.5,
    'CRCUREA':      1.5,
}
DEFAULT_BF_MULT = 1.5  # productos sin histórico BF


@bp.route('/api/planta/forecast-black-friday', methods=['GET'])
def forecast_black_friday():
    """Calcula stock extra necesario por SKU para Black Friday.

    Ventana pico: 14 días alrededor de BF (último viernes de noviembre).
    Multiplicador por SKU basado en BF 2025 (real).

    Devuelve:
      {
        bf_fecha: 'YYYY-MM-DD',
        ventana_pico: ['inicio', 'fin'],
        fecha_limite_stock: 'YYYY-MM-DD' (BF - 7d pipeline),
        skus: [{producto, vel_normal, vel_pico, mult, extra_unidades, extra_kg, lote_extra_kg}],
        kpis: total_lotes_extra_necesarios, total_kg_extra
      }

    Query params:
      year: año (default año actual o próximo si ya pasó BF)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    year = int(request.args.get('year') or datetime.now().year)
    # BF = último viernes de noviembre
    nov_30 = date(year, 11, 30)
    while nov_30.weekday() != 4:  # 4 = viernes
        nov_30 -= timedelta(days=1)
    bf_fecha = nov_30
    ventana_inicio = bf_fecha - timedelta(days=4)
    ventana_fin = bf_fecha + timedelta(days=10)
    fecha_limite_stock = bf_fecha - timedelta(days=7)  # pipeline 7d

    conn = get_db(); c = conn.cursor()

    # SKUs activos con ventas
    skus = c.execute("""
        SELECT spc.producto_nombre, fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
          AND COALESCE(spc.estado, 'activo') NOT IN ('descontinuado', 'pausado')
    """).fetchall()

    skus_out = []
    total_extra_kg = 0
    total_lotes_extra = 0

    for producto, lote_kg_default in skus:
        # Buscar SKUs Shopify mapeados → tomar el primero con multiplicador conocido
        skus_shop = c.execute("""
            SELECT sku FROM sku_producto_map
            WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
              AND COALESCE(activo, 1) = 1
        """, (producto,)).fetchall()
        mult = None
        sku_main = None
        for (s,) in skus_shop:
            if s in BF_MULTIPLIERS:
                mult = BF_MULTIPLIERS[s]
                sku_main = s
                break
        if mult is None:
            mult = DEFAULT_BF_MULT
            sku_main = skus_shop[0][0] if skus_shop else producto

        # Velocidad real reciente
        vel_real, ftend = _velocidad_total_producto(c, producto)
        vel_real_p = vel_real * (ftend or 1)
        if vel_real_p < 0.05: continue  # ignorar productos sin ventas

        vel_pico = vel_real_p * mult
        # Stock extra para 14 días de ventana pico
        extra_unidades = (vel_pico - vel_real_p) * 14  # solo el "delta" sobre normal
        if extra_unidades < 1: continue

        factor_g = _factor_g_por_unidad(c, producto)
        extra_kg = (extra_unidades * factor_g) / 1000

        lote_kg = lote_kg_default or 30
        # ¿Cuántos lotes EXTRA necesito?
        lotes_extra = max(1, round(extra_kg / lote_kg))

        skus_out.append({
            'producto': producto,
            'sku_main': sku_main,
            'velocidad_normal_u_d': round(vel_real_p, 2),
            'velocidad_pico_u_d': round(vel_pico, 2),
            'multiplicador': mult,
            'extra_unidades_pico': int(extra_unidades),
            'extra_kg_pico': round(extra_kg, 1),
            'lote_kg_actual': lote_kg,
            'lotes_extra_recomendados': lotes_extra,
            'urgencia': 'alta' if mult >= 4 else 'media' if mult >= 2.5 else 'baja',
        })
        total_extra_kg += extra_kg
        total_lotes_extra += lotes_extra

    # Ordenar por multiplicador descendente
    skus_out.sort(key=lambda x: -x['multiplicador'])

    return jsonify({
        'year': year,
        'bf_fecha': bf_fecha.isoformat(),
        'cyber_monday': (bf_fecha + timedelta(days=3)).isoformat(),
        'ventana_pico_inicio': ventana_inicio.isoformat(),
        'ventana_pico_fin': ventana_fin.isoformat(),
        'fecha_limite_stock': fecha_limite_stock.isoformat(),
        'fecha_inicio_pre_stock': (bf_fecha - timedelta(days=60)).isoformat(),
        'skus': skus_out,
        'kpis': {
            'total_skus_afectados': len(skus_out),
            'total_lotes_extra': total_lotes_extra,
            'total_kg_extra': round(total_extra_kg, 1),
        },
        'reglas': [
            'Multiplicadores empíricos basados en BF 2025 (calculado venta nov vs normal)',
            'Ventana pico = 4d antes BF + 10d después (incluye Cyber Monday)',
            'Pre-stock: producir lotes EXTRA en sept-oct para que estén listos antes de ' + fecha_limite_stock.isoformat(),
            'Pipeline 7d → fabricar máximo el ' + fecha_limite_stock.isoformat(),
            'BF 2025 hubo escasez · NO repetir',
        ],
    })


# NOTA: el endpoint para registrar lote real fue ELIMINADO porque era
# REDUNDANTE con /api/produccion (inventario.py) que ya descuenta MP con
# FEFO + transacción + anti-duplicado. El botón "✓ Hecho" del Centro de
# Acción ahora llama directamente /api/produccion (Sebastián 30-abr-2026).


# ════════════════════════════════════════════════════════════════════════
# AUTO-SC IA — generación automática de Solicitudes de Compra MP
# ════════════════════════════════════════════════════════════════════════
# Sebastián 30-abr-2026: "lo más automático posible. Primeros 5 días del
# mes genera órdenes para los 2 meses siguientes. Horizonte 60 y 90 días.
# Buffer IA analiza ventas para decir vamos aumentando. Va directo a
# Compras (Catalina + Alejandro). Email a Alejandro".

def _factor_buffer_ia(c, producto):
    """Buffer multiplicador basado en tendencia de ventas reales.

    Compara velocidad últimos 30d vs últimos 180d:
      crecimiento >50%: 1.50× (negocio explotando, pedir mucho más)
      crecimiento >25%: 1.30× (negocio creciendo fuerte)
      crecimiento >10%: 1.15× (crecimiento moderado)
      crecimiento  ±10%: 1.00× (estable)
      decrecimiento >10%: 0.90× (vendiendo menos)
    """
    try:
        # Últimos 30d
        v30 = _ventas_diarias_por_sku(c, producto, dias=30)
        v30_total = sum(q for _, q in v30)
        # Últimos 180d
        v180 = _ventas_diarias_por_sku(c, producto, dias=180)
        v180_total = sum(q for _, q in v180)
    except Exception:
        return 1.0

    vel_30 = v30_total / 30 if v30_total else 0
    vel_180 = v180_total / 180 if v180_total else 0
    if vel_180 < 0.05:
        return 1.0
    ratio = vel_30 / vel_180
    if ratio >= 1.50: return 1.50
    if ratio >= 1.25: return 1.30
    if ratio >= 1.10: return 1.15
    if ratio >= 0.90: return 1.00
    return 0.90


def _factor_buffer_estacional(fecha_objetivo):
    """Multiplicador por mes (anticipación BF + Día Madres + etc.)."""
    if fecha_objetivo is None:
        return 1.0
    mes = fecha_objetivo.month
    # Buffer pre-Black Friday (compra en sept-oct para llegar lista a nov)
    if mes in (9, 10):
        return 1.20  # +20% pre-BF
    if mes == 11:
        return 1.50  # +50% el mes BF (BF + Cyber Monday)
    if mes == 5:
        return 1.10  # +10% Día Madres
    if mes == 12:
        return 1.20  # +20% Navidad / cierre año
    return 1.0


def _calcular_auto_sc(conn, horizontes_dias=(60, 90), modo='mensual'):
    """Calcula las SC que deberían crearse automáticamente.

    Args:
      horizontes_dias: tupla con horizontes a evaluar (default 60 y 90)
      modo: 'mensual' (cron 1-5 del mes) o 'urgente' (cron semanal lunes,
            solo MPs con stockout en próximos 14d)

    Returns:
      {
        scs_por_proveedor: {proveedor: [items]},
        items_huerfanos: [items sin proveedor sugerido],
        modo, horizontes_dias, fecha_analisis,
        kpis: {total_items, total_proveedores, total_g, mp_criticas}
      }
    """
    c = conn.cursor()
    fecha_hoy = datetime.now().date()

    try:
        from blueprints.programacion import _get_mp_stock, _norm_mp_name
        mp_stock_map = _get_mp_stock(conn)
    except Exception:
        mp_stock_map = {}
        _norm_mp_name = lambda x: str(x or '').upper().strip()

    skus = c.execute("""
        SELECT spc.producto_nombre, fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1 AND COALESCE(spc.estado, 'activo') NOT IN ('descontinuado', 'pausado')
    """).fetchall()
    sku_lote = {p: l for p, l in skus}

    # Cache fórmulas + buffer IA por producto
    formulas_cache = {}
    buffer_ia_cache = {}

    def _get_formula(prod):
        if prod not in formulas_cache:
            rows = c.execute("""
                SELECT material_id, material_nombre, porcentaje
                FROM formula_items WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
            """, (prod,)).fetchall()
            formulas_cache[prod] = rows
        return formulas_cache[prod]

    def _get_buffer_ia(prod):
        if prod not in buffer_ia_cache:
            buffer_ia_cache[prod] = _factor_buffer_ia(c, prod)
        return buffer_ia_cache[prod]

    def _stock_de(mat_id, mat_nombre):
        mid = str(mat_id or '').strip()
        if mid in mp_stock_map: return float(mp_stock_map[mid])
        if mid.upper() in mp_stock_map: return float(mp_stock_map[mid.upper()])
        nom_up = str(mat_nombre or '').strip().upper()
        if nom_up and nom_up in mp_stock_map: return float(mp_stock_map[nom_up])
        try:
            nom_norm = _norm_mp_name(mat_nombre or '')
            if nom_norm and nom_norm in mp_stock_map: return float(mp_stock_map[nom_norm])
        except Exception: pass
        return 0

    # Eventos del Calendar en el horizonte mayor
    eventos = _calendar_events_cached()
    horizonte_max = max(horizontes_dias)
    fecha_limite = fecha_hoy + timedelta(days=horizonte_max)

    producciones = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if f < fecha_hoy or f > fecha_limite:
            continue
        producto_match = None
        for prod_nom in sku_lote.keys():
            alias = _alias_calendar_for(c, prod_nom)
            score = _match_producto_evento(prod_nom, alias, ev.get('titulo'), ev.get('descripcion', ''))
            if score >= 60:
                producto_match = prod_nom
                break
        if not producto_match:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', '')) or sku_lote.get(producto_match) or 30
        producciones.append({'fecha': f, 'producto': producto_match, 'kg': kg})
    producciones.sort(key=lambda p: p['fecha'])

    # Acumular consumo por MP (con buffer IA por producto)
    mp_consumo = {}  # (mat_id, mat_nom) → {stock, consumos: [{fecha, g_neto, g_ajustado}]}
    for prod_evento in producciones:
        items = _get_formula(prod_evento['producto'])
        if not items: continue
        kg = prod_evento['kg']
        total_g = kg * 1000
        buf_ia = _get_buffer_ia(prod_evento['producto'])
        buf_est = _factor_buffer_estacional(prod_evento['fecha'])
        for mat_id, mat_nom, pct in items:
            try: pct_f = float(pct or 0)
            except Exception: pct_f = 0
            req_g_neto = total_g * pct_f / 100
            if req_g_neto <= 0: continue
            req_g_ajustado = req_g_neto * buf_ia * buf_est
            key = (mat_id, mat_nom or mat_id)
            if key not in mp_consumo:
                mp_consumo[key] = {
                    'material_id': mat_id,
                    'material_nombre': mat_nom or mat_id,
                    'stock_g': _stock_de(mat_id, mat_nom),
                    'consumos': [],
                }
            mp_consumo[key]['consumos'].append({
                'fecha': prod_evento['fecha'],
                'producto': prod_evento['producto'],
                'g_neto': req_g_neto,
                'g_ajustado': req_g_ajustado,
                'buf_ia': buf_ia,
                'buf_est': buf_est,
            })

    # Determinar qué MPs requieren SC
    items_a_pedir = []
    for key, info in mp_consumo.items():
        # Saldo proyectado a 60d y 90d
        saldos_por_horizonte = {}
        for h in horizontes_dias:
            fecha_h = fecha_hoy + timedelta(days=h)
            consumo_h = sum(c['g_ajustado'] for c in info['consumos'] if c['fecha'] <= fecha_h)
            saldo_h = info['stock_g'] - consumo_h
            saldos_por_horizonte[h] = saldo_h

        # Modo urgente: solo MPs con stockout en próximos 14 días
        if modo == 'urgente':
            consumo_14d = sum(c['g_ajustado'] for c in info['consumos'] if (c['fecha'] - fecha_hoy).days <= 14)
            saldo_14d = info['stock_g'] - consumo_14d
            if saldo_14d >= 0:
                continue  # no urgente
            # Cantidad a pedir: lo que falta + 30d cobertura
            consumo_30d = sum(c['g_ajustado'] for c in info['consumos'] if (c['fecha'] - fecha_hoy).days <= 30)
            cantidad = abs(saldo_14d) + consumo_30d
        else:
            # Modo mensual: pedir para llegar a saldo positivo en horizonte 60d y mantener cobertura 90d
            saldo_60 = saldos_por_horizonte.get(60, 0)
            saldo_90 = saldos_por_horizonte.get(90, 0)
            if saldo_60 >= 0 and saldo_90 >= 0:
                continue  # alcanza para ambos horizontes
            # Cantidad: déficit del horizonte 90d (ya incluye buffers IA + estacional)
            cantidad = abs(min(saldo_60, saldo_90))
            # Más buffer cobertura adicional 30d
            consumo_30d_post = sum(c['g_ajustado'] for c in info['consumos'] if (c['fecha'] - fecha_hoy).days <= 30)
            cantidad += consumo_30d_post * 0.5  # +15d cobertura extra

        if cantidad < 50:  # ignorar MPs con cantidad mínima absurda (<50g)
            continue

        # Lead time + proveedor
        lt = c.execute("""
            SELECT lead_time_dias, origen, proveedor_principal
            FROM mp_lead_time_config WHERE material_id=?
        """, (info['material_id'],)).fetchone()
        lead = lt[0] if lt else 14
        origen = lt[1] if lt else 'local'
        proveedor = (lt[2] if lt else '') or 'Sin proveedor sugerido'

        # Buffer IA promedio (de los productos que la usan)
        bufs_ia = [c['buf_ia'] for c in info['consumos']]
        buf_ia_avg = sum(bufs_ia) / len(bufs_ia) if bufs_ia else 1.0

        # Justificación
        productos_que_la_usan = list(set(c['producto'] for c in info['consumos']))[:5]
        justif = (
            f'Auto-SC IA · stock {int(info["stock_g"])}g · '
            f'consumirán {int(sum(c["g_ajustado"] for c in info["consumos"]))}g en {horizonte_max}d · '
            f'usada por {len(productos_que_la_usan)} producto(s): {", ".join(productos_que_la_usan[:3])} · '
            f'buffer IA tendencia ×{buf_ia_avg:.2f} · lead {lead}d ({origen})'
        )

        items_a_pedir.append({
            'material_id': info['material_id'],
            'material_nombre': info['material_nombre'],
            'cantidad_g': round(cantidad, 0),
            'unidad': 'g',
            'justificacion': justif,
            'proveedor_sugerido': proveedor,
            'lead_time_dias': lead,
            'origen': origen,
            'stock_actual_g': round(info['stock_g'], 0),
            'consumo_total_g': round(sum(c['g_ajustado'] for c in info['consumos']), 0),
            'buffer_ia_avg': round(buf_ia_avg, 2),
            'productos_que_la_usan': productos_que_la_usan,
            'saldo_60d': round(saldos_por_horizonte.get(60, 0), 0),
            'saldo_90d': round(saldos_por_horizonte.get(90, 0), 0),
        })

    # Agrupar por proveedor → 1 SC por proveedor
    scs_por_proveedor = {}
    huerfanos = []
    for item in items_a_pedir:
        prov = item['proveedor_sugerido']
        if not prov or prov == 'Sin proveedor sugerido':
            huerfanos.append(item)
        else:
            scs_por_proveedor.setdefault(prov, []).append(item)

    return {
        'modo': modo,
        'horizontes_dias': list(horizontes_dias),
        'fecha_analisis': datetime.now().isoformat(),
        'producciones_analizadas': len(producciones),
        'scs_por_proveedor': scs_por_proveedor,
        'items_huerfanos': huerfanos,
        'kpis': {
            'total_items': len(items_a_pedir),
            'total_proveedores': len(scs_por_proveedor),
            'total_g': round(sum(i['cantidad_g'] for i in items_a_pedir), 0),
            'mp_huerfanas': len(huerfanos),
        },
    }


@bp.route('/api/planta/auto-sc-preview', methods=['GET'])
def auto_sc_preview():
    """Devuelve qué SCs se generarían SIN crearlas (para revisar).

    Query params:
      modo: 'mensual' (default) o 'urgente'
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    modo = (request.args.get('modo') or 'mensual').strip()
    conn = get_db()
    return jsonify(_calcular_auto_sc(conn, horizontes_dias=(60, 90), modo=modo))


@bp.route('/api/planta/auto-sc-generar', methods=['POST'])
def auto_sc_generar():
    """Crea las SCs automáticamente. Acepta:
      - sesión 'compras_user' (manual)
      - ?clave=AUTO_PLAN_CRON_KEY (cron externo Render)

    Body:
      modo: 'mensual' (default) o 'urgente'
      enviar_email: bool (default true)
    """
    clave_cron = request.args.get('clave', '') or (request.json or {}).get('clave', '')
    clave_esperada = os.environ.get('AUTO_PLAN_CRON_KEY', '')
    es_cron = bool(clave_esperada and clave_cron == clave_esperada)
    if not es_cron and 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    d = request.json or {}
    modo = (d.get('modo') or request.args.get('modo') or 'mensual').strip()
    enviar_email = d.get('enviar_email', True)
    user = 'auto-plan-ia' if es_cron else session.get('compras_user', 'manual')

    conn = get_db(); c = conn.cursor()
    plan = _calcular_auto_sc(conn, horizontes_dias=(60, 90), modo=modo)

    # Crear las SCs reales en estado 'Pendiente' (Sebastián: va directo a
    # Compras; Catalina + Alejandro revisan ahí)
    scs_creadas = []
    fecha_hoy_iso = datetime.now().date().isoformat()
    for proveedor, items in plan['scs_por_proveedor'].items():
        # Generar número
        c.execute("""
            SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)), 0)
            FROM solicitudes_compra WHERE numero LIKE ?
        """, (f"SOL-{datetime.now().strftime('%Y')}-%",))
        n = (c.fetchone()[0] or 0) + 1
        numero = f"SOL-{datetime.now().strftime('%Y')}-{n:04d}"

        observ = f'🤖 Auto-SC IA ({modo}) · proveedor: {proveedor} · horizonte 60-90d · buffers IA tendencia + estacional aplicados'

        c.execute("""
            INSERT INTO solicitudes_compra
            (numero, fecha, estado, solicitante, urgencia, observaciones, area, empresa, categoria, tipo, fecha_requerida, valor)
            VALUES (?, ?, 'Pendiente', ?, ?, ?, 'Produccion', 'Espagiria', 'Materia Prima', 'Compra', ?, 0)
        """, (numero, datetime.now().isoformat(), user,
              'Alta' if modo == 'urgente' else 'Normal',
              observ, fecha_hoy_iso))

        for it in items:
            try:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                    (numero, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion, valor_estimado, proveedor_sugerido)
                    VALUES (?, ?, ?, ?, 'g', ?, 0, ?)
                """, (numero, it['material_id'], it['material_nombre'],
                      it['cantidad_g'], it['justificacion'], it['proveedor_sugerido']))
            except sqlite3.OperationalError:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                    (numero, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion, valor_estimado)
                    VALUES (?, ?, ?, ?, 'g', ?, 0)
                """, (numero, it['material_id'], it['material_nombre'],
                      it['cantidad_g'], it['justificacion']))

        scs_creadas.append({
            'numero': numero,
            'proveedor': proveedor,
            'items_count': len(items),
            'total_g': round(sum(it['cantidad_g'] for it in items), 0),
        })

    conn.commit()

    # Email a Alejandro con resumen (desde email_destinatarios_config)
    email_enviado = False
    if enviar_email and scs_creadas:
        try:
            html = _generar_html_auto_sc(plan, scs_creadas, modo)
            # Sebastián: "email a alejandro". Intentamos varias rutas:
            # 1) rol gerencia_produccion (suele ser Alejandro)
            # 2) email LIKE %alejandro%
            # 3) flag recibe_compras_aprob = 1 (compras flow)
            destinatarios = []
            try:
                rows = c.execute("""
                    SELECT email FROM email_destinatarios_config
                    WHERE activo=1 AND email != ''
                      AND (rol='gerencia_produccion' OR LOWER(email) LIKE '%alejandro%' OR recibe_compras_aprob=1)
                """).fetchall()
                destinatarios = [r[0] for r in rows if r[0]]
            except Exception:
                destinatarios = []
            if not destinatarios:
                try:
                    rows = c.execute("""
                        SELECT email FROM email_destinatarios_config
                        WHERE activo=1 AND email != ''
                    """).fetchall()
                    destinatarios = [r[0] for r in rows if r[0]]
                except Exception:
                    destinatarios = []
            if destinatarios:
                import threading, sys, os as _os
                sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
                from notificaciones import SistemaNotificaciones
                notif = SistemaNotificaciones()
                threading.Thread(
                    target=notif._enviar_email,
                    args=(f'🤖 Auto-SC IA · {len(scs_creadas)} SCs creadas ({modo})', html, destinatarios),
                    daemon=True
                ).start()
                email_enviado = True
        except Exception as e:
            log.warning(f'Email auto-SC fallo: {e}')

    # Log en auto_plan_runs (esquema real: ejecutado_por, compras_creadas, error, payload_json)
    try:
        notas = f'{plan["kpis"]["total_items"]} items · {len(scs_creadas)} SCs · email={email_enviado}'
        c.execute("""
            INSERT INTO auto_plan_runs
              (ejecutado_at, ejecutado_por, tipo, horizonte_dias,
               producciones_creadas, compras_creadas, alertas_criticas,
               emails_enviados, error, payload_json)
            VALUES (?, ?, ?, ?, 0, ?, 0, ?, NULL, ?)
        """, (
            datetime.now().isoformat(), user, f'auto_sc_{modo}',
            90 if modo == 'mensual' else 14,
            len(scs_creadas),
            1 if email_enviado else 0,
            json.dumps({'modo': modo, 'kpis': plan['kpis'],
                        'scs': scs_creadas, 'notas': notas}, default=str),
        ))
        conn.commit()
    except Exception as e:
        log.warning(f'Log auto-sc fallo: {e}')

    return jsonify({
        'ok': True,
        'modo': modo,
        'scs_creadas': scs_creadas,
        'items_huerfanos': plan['items_huerfanos'],
        'email_enviado': email_enviado,
        'kpis': plan['kpis'],
        'mensaje': f'✅ {len(scs_creadas)} SCs creadas en estado Pendiente · {plan["kpis"]["total_items"]} MPs · email={"sí" if email_enviado else "no"}',
    })


def _generar_html_auto_sc(plan, scs_creadas, modo):
    """HTML email para Alejandro con resumen de las SCs creadas."""
    fecha_hoy = datetime.now().date()
    titulo = '🚨 SC Urgentes (lunes)' if modo == 'urgente' else '🤖 SC Mensuales 60-90d'
    bgHeader = '#dc2626' if modo == 'urgente' else '#0891b2'

    html_scs = ''
    for sc in scs_creadas:
        items = plan['scs_por_proveedor'].get(sc['proveedor'], [])
        html_scs += f'<div style="background:#f8fafc;border-left:4px solid {bgHeader};padding:10px 14px;margin-bottom:8px;border-radius:0 6px 6px 0">'
        html_scs += f'<b>{sc["numero"]}</b> · proveedor: <b>{sc["proveedor"]}</b> · {sc["items_count"]} MPs · {int(sc["total_g"]/1000)} kg total'
        html_scs += '<ul style="margin:6px 0 0 18px;padding:0;font-size:11px;color:#475569">'
        for it in items[:6]:
            html_scs += f'<li><b>{it["material_nombre"]}</b>: {int(it["cantidad_g"]/1000)} kg (buf IA ×{it["buffer_ia_avg"]})</li>'
        if len(items) > 6:
            html_scs += f'<li style="color:#94a3b8">+ {len(items)-6} MPs más</li>'
        html_scs += '</ul></div>'

    huerfanos_html = ''
    if plan['items_huerfanos']:
        huerfanos_html = '<div style="background:#fef3c7;border:1px solid #fcd34d;padding:10px;border-radius:6px;margin-top:14px">'
        huerfanos_html += f'<b style="color:#92400e">⚠️ {len(plan["items_huerfanos"])} MPs SIN proveedor sugerido</b>'
        huerfanos_html += '<div style="font-size:11px;color:#78350f;margin-top:4px">No se crearon SCs porque mp_lead_time_config no tiene proveedor_principal:</div>'
        huerfanos_html += '<ul style="margin:4px 0 0 18px;padding:0;font-size:11px;color:#78350f">'
        for it in plan['items_huerfanos'][:8]:
            huerfanos_html += f'<li>{it["material_nombre"]}: {int(it["cantidad_g"]/1000)} kg</li>'
        huerfanos_html += '</ul></div>'

    return f'''<!DOCTYPE html>
<html><body style="font-family:-apple-system,sans-serif;background:#f3f4f6;padding:20px;margin:0">
  <div style="max-width:640px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,.08)">
    <div style="background:{bgHeader};color:#fff;padding:20px">
      <h2 style="margin:0;font-size:20px">{titulo}</h2>
      <p style="margin:4px 0 0;opacity:.9;font-size:13px">{fecha_hoy.strftime("%d-%b-%Y")} · Auto-Plan IA</p>
    </div>
    <div style="padding:20px">
      <div style="background:#ecfdf5;border:1px solid #6ee7b7;padding:10px;border-radius:6px;margin-bottom:14px;font-size:13px;color:#065f46">
        ✅ <b>{len(scs_creadas)} solicitudes creadas</b> · {plan["kpis"]["total_items"]} MPs · {plan["kpis"]["total_g"]/1000:.0f} kg total<br>
        Estado: <b>Pendiente</b> · Catalina y Alejandro pueden revisar y aprobar en <a href="/solicitudes" style="color:#065f46">/solicitudes</a>
      </div>
      <h3 style="color:#0f172a;font-size:14px;margin:14px 0 8px">📋 Solicitudes creadas</h3>
      {html_scs}
      {huerfanos_html}
      <div style="margin-top:18px;padding:12px;background:#f1f5f9;border-radius:8px;font-size:11px;color:#64748b">
        Buffer IA aplicado: tendencia (vel 30d vs 180d) + estacional (BF/Madres/Navidad).<br>
        Auto-SC IA · cron mensual día 1-5 (60-90d) + cron lunes (urgentes 14d).
      </div>
    </div>
  </div>
</body></html>'''


@bp.route('/api/planta/auto-sc-status', methods=['GET'])
def auto_sc_status():
    """Status compacto para el panel Auto-SC IA en el tab Plan.

    Devuelve último run (mensual y urgente), SCs creadas en el mes actual y
    en el anterior, próximas ventanas (siguiente lunes y siguiente día-1),
    y la lista de las últimas 10 SCs creadas por la IA.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from datetime import date, timedelta
    conn = get_db(); c = conn.cursor()
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)
    if inicio_mes.month == 1:
        inicio_mes_pasado = inicio_mes.replace(year=inicio_mes.year - 1, month=12)
    else:
        inicio_mes_pasado = inicio_mes.replace(month=inicio_mes.month - 1)

    def _last_run(tipo):
        try:
            row = c.execute("""
                SELECT ejecutado_at, compras_creadas, payload_json, error, emails_enviados
                FROM auto_plan_runs
                WHERE tipo = ?
                ORDER BY id DESC LIMIT 1
            """, (tipo,)).fetchone()
            if not row:
                return None
            return {
                'ejecutado_at': row[0],
                'scs_creadas': row[1] or 0,
                'payload_json': row[2] or '',
                'error': row[3] or '',
                'emails_enviados': row[4] or 0,
            }
        except Exception:
            return None

    def _count_scs_desde(d_iso):
        try:
            row = c.execute("""
                SELECT COUNT(*) FROM solicitudes_compra
                WHERE solicitante = 'auto-plan-ia' AND date(fecha) >= ?
            """, (d_iso,)).fetchone()
            return row[0] if row else 0
        except Exception:
            return 0

    def _proximo_lunes():
        dias = (7 - hoy.weekday()) % 7
        if dias == 0:
            dias = 7
        return hoy + timedelta(days=dias)

    def _proxima_ventana_mensual():
        if hoy.day <= 5:
            return hoy
        if hoy.month == 12:
            return date(hoy.year + 1, 1, 1)
        return date(hoy.year, hoy.month + 1, 1)

    recientes = []
    try:
        rows = c.execute("""
            SELECT s.id, s.numero, s.estado, s.fecha, s.observaciones, s.valor,
                   (SELECT COUNT(*) FROM solicitudes_compra_items i WHERE i.numero = s.numero) AS items_count,
                   (SELECT GROUP_CONCAT(DISTINCT proveedor_sugerido)
                      FROM solicitudes_compra_items i WHERE i.numero = s.numero) AS proveedor
            FROM solicitudes_compra s
            WHERE s.solicitante = 'auto-plan-ia'
            ORDER BY s.id DESC LIMIT 12
        """).fetchall()
        cols = [d[0] for d in c.description]
        recientes = [dict(zip(cols, r)) for r in rows]
    except Exception:
        recientes = []

    return jsonify({
        'hoy': hoy.isoformat(),
        'dia_mes': hoy.day,
        'ventana_mensual_activa': hoy.day <= 5,
        'proximo_lunes': _proximo_lunes().isoformat(),
        'proxima_ventana_mensual': _proxima_ventana_mensual().isoformat(),
        'last_mensual': _last_run('auto_sc_mensual'),
        'last_urgente': _last_run('auto_sc_urgente'),
        'scs_mes_actual': _count_scs_desde(inicio_mes.isoformat()),
        'scs_mes_pasado': _count_scs_desde(inicio_mes_pasado.isoformat()) - _count_scs_desde(inicio_mes.isoformat()),
        'recientes': recientes,
    })


# ════════════════════════════════════════════════════════════════════════
# AUTO-SC IA · MEE (Material de Empaque y Etiquetas)
# ════════════════════════════════════════════════════════════════════════
# Sebastian (1-may-2026): "envases China 9m, etiquetas las pedimos al
# envasar, serigrafía 20d antes, plegadiza no aplica, el resto local 60-90d
# como MP". Espejo del Auto-SC de MP pero proyectando ventas por SKU →
# componentes MEE, agrupando por (proveedor, origen).
#
# Horizontes:
#   · China:  270d (9m) — lead time 180d + buffer producción 90d
#   · Local:  90d  — espejo del MP local
#   · Urgente: 30d — solo lo que stockout en próximas 4 semanas
#
# Filtros:
#   · sku_mee_config.aplica = 0 → ignorado (plegadiza)
#   · mee_lead_time_config.disparo_d20 = 1 → ignorado (serigrafía/tampografía
#     van por cron diario D-20, fase aparte)


def _calcular_auto_sc_mee(conn, modo='mensual', origen_filtro=None, generico=False):
    """Calcula SCs MEE proyectando ventas por SKU.

    Args:
      modo: 'mensual' (default, horizonte por origen) o 'urgente' (30d todos)
      origen_filtro: 'China'|'Local'|None (None = todos)
      generico: si True, para SKUs SIN mapping en sku_mee_config genera items
                "abiertos" tipo envase/tapa/etiqueta sin código MEE específico.
                Catalina los asigna en Compras y el sistema aprende.
                Sebastián (1-may-2026): "el sistema ya sabe envase que necesita,
                que se solicite automatico, y sea catalina quien asigne".

    Returns:
      {scs_por_proveedor (clave='Proveedor (Origen)'), items_huerfanos,
       items_genericos, kpis, fecha_analisis, modo, origen_filtro}
    """
    c = conn.cursor()
    fecha_hoy = datetime.now().date()

    HORIZONTE_CHINA = 270
    HORIZONTE_LOCAL = 90
    HORIZONTE_URGENTE = 30

    # 1) Mappings sku → componentes MEE (con cantidad_por_unidad)
    sku_components = {}
    try:
        rows = c.execute("""
            SELECT sku_codigo, mee_codigo, componente_tipo, cantidad_por_unidad
            FROM sku_mee_config
            WHERE aplica = 1
        """).fetchall()
        for sku, mee, tipo, cant in rows:
            sku_components.setdefault(sku, []).append({
                'mee_codigo': mee,
                'tipo': tipo,
                'cantidad_por_unidad': float(cant or 1),
            })
    except sqlite3.OperationalError:
        return _empty_auto_sc_mee_result(modo, origen_filtro, fecha_hoy,
                                          razon='sku_mee_config no existe')

    # 2) Config MEE (lead time, MOQ, proveedor, origen)
    # Filtros: aplica=1, NOT disparo_d20 (van por cron diario), NOT
    # disparo_post_envasado (etiquetas se piden al envasar, alerta aparte)
    mee_config = {}
    try:
        try:
            rows = c.execute("""
                SELECT mee_codigo, proveedor_principal, origen, lead_time_dias,
                       moq_unidades, precio_unit, disparo_d20,
                       COALESCE(disparo_post_envasado, 0), aplica
                FROM mee_lead_time_config
                WHERE aplica = 1
                  AND COALESCE(disparo_d20, 0) = 0
                  AND COALESCE(disparo_post_envasado, 0) = 0
            """).fetchall()
        except sqlite3.OperationalError:
            # Esquema legacy sin disparo_post_envasado
            rows = c.execute("""
                SELECT mee_codigo, proveedor_principal, origen, lead_time_dias,
                       moq_unidades, precio_unit, disparo_d20, 0, aplica
                FROM mee_lead_time_config
                WHERE aplica = 1 AND COALESCE(disparo_d20, 0) = 0
            """).fetchall()
        for mc, prov, ori, lt, moq, prec, d20, post_env, ap in rows:
            if origen_filtro and ori != origen_filtro:
                continue
            mee_config[mc] = {
                'proveedor': (prov or '').strip() or '(sin proveedor)',
                'origen': ori or 'Local',
                'lead_time_dias': int(lt or 30),
                'moq_unidades': int(moq or 0),
                'precio_unit': float(prec or 0),
            }
    except sqlite3.OperationalError:
        return _empty_auto_sc_mee_result(modo, origen_filtro, fecha_hoy,
                                          razon='mee_lead_time_config no existe')

    if not mee_config:
        return _empty_auto_sc_mee_result(modo, origen_filtro, fecha_hoy,
                                          razon='Ningún MEE configurado en mee_lead_time_config (origen/proveedor)')

    # 3) Stock MEE
    stock_mee = {}
    try:
        for cod, st in c.execute("SELECT codigo, COALESCE(stock_actual,0) FROM maestro_mee").fetchall():
            stock_mee[cod] = float(st or 0)
    except Exception:
        pass

    # 4) SKUs activos
    skus_lista = []
    try:
        rows = c.execute("""
            SELECT producto_nombre FROM sku_planeacion_config
            WHERE activo = 1
              AND COALESCE(estado, 'activo') NOT IN ('descontinuado', 'pausado', 'sin_ventas')
        """).fetchall()
        skus_lista = [r[0] for r in rows]
    except Exception:
        pass

    # 5) Proyección demanda por MEE (sumar todos los SKUs que la usan)
    demanda_mee = {}  # mee_codigo → {cantidad_total, justificaciones[]}
    for sku in skus_lista:
        if sku not in sku_components:
            continue
        try:
            v180 = _ventas_diarias_por_sku(c, sku, dias=180)
            v180_total = sum(q for _, q in v180)
        except Exception:
            v180_total = 0
        vel_diaria = v180_total / 180.0 if v180_total else 0
        if vel_diaria <= 0:
            continue
        try:
            buf_ia = _factor_buffer_ia(c, sku)
        except Exception:
            buf_ia = 1.0
        for comp in sku_components[sku]:
            mee_cod = comp['mee_codigo']
            if mee_cod not in mee_config:
                continue
            cfg = mee_config[mee_cod]
            origen = cfg['origen']
            if modo == 'urgente':
                horizonte = HORIZONTE_URGENTE
            elif origen == 'China':
                horizonte = HORIZONTE_CHINA
            else:
                horizonte = HORIZONTE_LOCAL
            fecha_medio = fecha_hoy + timedelta(days=horizonte // 2)
            buf_est = _factor_buffer_estacional(fecha_medio)
            unidades_sku = vel_diaria * horizonte * buf_ia * buf_est
            cant_mee = unidades_sku * comp['cantidad_por_unidad']
            if mee_cod not in demanda_mee:
                demanda_mee[mee_cod] = {
                    'cantidad_total': 0,
                    'horizonte': horizonte,
                    'justificaciones': [],
                }
            demanda_mee[mee_cod]['cantidad_total'] += cant_mee
            demanda_mee[mee_cod]['justificaciones'].append({
                'sku': sku,
                'velocidad_diaria': round(vel_diaria, 2),
                'horizonte_dias': horizonte,
                'buf_ia': buf_ia,
                'buf_est': round(buf_est, 2),
                'cantidad_por_unidad': comp['cantidad_por_unidad'],
                'tipo_componente': comp['tipo'],
                'unidades_mee_estimadas': round(cant_mee, 0),
            })

    # 7) MODO GENÉRICO: para SKUs sin mapping, generar items "abiertos"
    # tipo envase + tapa + etiqueta + (serigrafia si aplica) que Catalina
    # asignará en Compras. El sistema aprende del mapeo (PUT sc-mee-asignar).
    items_genericos = []
    if generico:
        for sku in skus_lista:
            # Si ya tiene mappings, NO generar genéricos (los reales toman precedencia)
            if sku in sku_components:
                continue
            try:
                v180 = _ventas_diarias_por_sku(c, sku, dias=180)
                v180_total = sum(q for _, q in v180)
            except Exception:
                v180_total = 0
            vel_diaria = v180_total / 180.0 if v180_total else 0
            if vel_diaria <= 0:
                continue
            try:
                buf_ia = _factor_buffer_ia(c, sku)
            except Exception:
                buf_ia = 1.0
            # Horizonte por default Local 90d (Catalina ajusta luego)
            horizonte = HORIZONTE_URGENTE if modo == 'urgente' else HORIZONTE_LOCAL
            fecha_medio = fecha_hoy + timedelta(days=horizonte // 2)
            buf_est = _factor_buffer_estacional(fecha_medio)
            cant_estimada = round(vel_diaria * horizonte * buf_ia * buf_est, 0)
            if cant_estimada <= 0:
                continue
            # Componentes "estándar" que necesita TODO SKU vendido
            componentes_default = [
                ('envase', '📦 Envase primario'),
                ('tapa', '🔘 Tapa o sistema dosificador'),
                ('etiqueta', '🏷️ Etiqueta del producto'),
            ]
            for tipo, label in componentes_default:
                items_genericos.append({
                    'sku_codigo': sku,
                    'componente_tipo': tipo,
                    'descripcion': f'{label} para {sku}',
                    'cantidad_unidades': cant_estimada,
                    'velocidad_diaria': round(vel_diaria, 2),
                    'horizonte_dias': horizonte,
                    'buf_ia': buf_ia,
                    'buf_est': round(buf_est, 2),
                    'mee_codigo': '',  # vacío = por asignar
                    'proveedor_sugerido': '',  # vacío = por asignar
                    'justificacion': (
                        f'🎯 GENÉRICO · {label} para {sku} · '
                        f'{cant_estimada:.0f} ud (vel {vel_diaria:.2f}/d × {horizonte}d × '
                        f'IA {buf_ia} × est {buf_est:.2f}) · '
                        f'CATALINA: asignar código MEE específico + proveedor; '
                        f'el sistema guardará el mapping para futuras SCs.'
                    ),
                })

    # 6) Déficit + MOQ + agrupar por (proveedor, origen)
    scs_por_proveedor = {}
    items_huerfanos = []
    for mee_cod, dem in demanda_mee.items():
        cant_demanda = dem['cantidad_total']
        stock = stock_mee.get(mee_cod, 0)
        deficit = max(0, cant_demanda - stock)
        if deficit <= 0:
            continue
        cfg = mee_config[mee_cod]
        moq = cfg['moq_unidades']
        cant_a_pedir = max(deficit, moq) if moq > 0 else deficit
        prov = cfg['proveedor']
        try:
            row = c.execute("SELECT descripcion FROM maestro_mee WHERE codigo=?", (mee_cod,)).fetchone()
            nombre = (row[0] if row else mee_cod) or mee_cod
        except Exception:
            nombre = mee_cod
        item = {
            'mee_codigo': mee_cod,
            'mee_nombre': nombre,
            'cantidad_unidades': round(cant_a_pedir, 0),
            'demanda_estimada': round(cant_demanda, 0),
            'stock_actual': round(stock, 0),
            'deficit': round(deficit, 0),
            'moq_aplicado': moq,
            'cobertura_dias': dem['horizonte'],
            'origen': cfg['origen'],
            'lead_time_dias': cfg['lead_time_dias'],
            'precio_unit': cfg['precio_unit'],
            'valor_estimado': round(cant_a_pedir * cfg['precio_unit'], 2) if cfg['precio_unit'] else 0,
            'justificaciones': dem['justificaciones'][:5],
            'proveedor_sugerido': prov,
        }
        # Justificación texto
        moq_msg = f' · MOQ {moq}' if moq > 0 and cant_a_pedir == moq else ''
        item['justificacion'] = (
            f"MEE {mee_cod} ({cfg['origen']}, lead {cfg['lead_time_dias']}d) · "
            f"stock {round(stock,0)} ud · demanda {round(cant_demanda,0)} ud "
            f"({dem['horizonte']}d) · déficit {round(deficit,0)}{moq_msg} · "
            f"pedir {round(cant_a_pedir,0)} ud"
        )
        if not prov or prov == '(sin proveedor)':
            items_huerfanos.append(item)
            continue
        clave = f"{prov} ({cfg['origen']})"
        scs_por_proveedor.setdefault(clave, []).append(item)

    total_items = sum(len(v) for v in scs_por_proveedor.values())
    total_unidades = sum(it['cantidad_unidades'] for items in scs_por_proveedor.values() for it in items)
    total_valor = sum(it['valor_estimado'] for items in scs_por_proveedor.values() for it in items)

    return {
        'modo': modo,
        'generico': generico,
        'origen_filtro': origen_filtro,
        'fecha_analisis': fecha_hoy.isoformat(),
        'scs_por_proveedor': scs_por_proveedor,
        'items_huerfanos': items_huerfanos,
        'items_genericos': items_genericos,
        'kpis': {
            'total_items': total_items,
            'total_proveedores': len(scs_por_proveedor),
            'total_unidades': round(total_unidades, 0),
            'total_valor_estimado': round(total_valor, 2),
            'mee_huerfanas': len(items_huerfanos),
            'skus_evaluados': len(skus_lista),
            'mee_evaluados': len(demanda_mee),
            'items_genericos_pendientes': len(items_genericos),
            'skus_genericos': len(set(it['sku_codigo'] for it in items_genericos)),
        },
    }


def _empty_auto_sc_mee_result(modo, origen_filtro, fecha_hoy, razon=''):
    """Resultado vacío con mensaje de razón (config faltante)."""
    return {
        'modo': modo,
        'origen_filtro': origen_filtro,
        'fecha_analisis': fecha_hoy.isoformat(),
        'scs_por_proveedor': {},
        'items_huerfanos': [],
        'items_genericos': [],
        'kpis': {
            'total_items': 0, 'total_proveedores': 0, 'total_unidades': 0,
            'total_valor_estimado': 0, 'mee_huerfanas': 0,
            'skus_evaluados': 0, 'mee_evaluados': 0,
            'items_genericos_pendientes': 0, 'skus_genericos': 0,
        },
        'razon_vacio': razon,
    }


@bp.route('/api/planta/auto-sc-mee-preview', methods=['GET'])
def auto_sc_mee_preview():
    """Preview Auto-SC MEE sin crear SCs reales.

    Query: ?modo=mensual|urgente  ?origen=China|Local  ?generico=1
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    modo = (request.args.get('modo') or 'mensual').strip()
    origen = (request.args.get('origen') or '').strip() or None
    generico = request.args.get('generico', '0') in ('1','true','yes')
    if origen and origen not in ('China', 'Local', 'Mixto'):
        return jsonify({'error': 'origen debe ser China, Local o Mixto'}), 400
    return jsonify(_calcular_auto_sc_mee(get_db(), modo=modo,
                                           origen_filtro=origen,
                                           generico=generico))


@bp.route('/api/planta/auto-sc-mee-generar', methods=['POST'])
def auto_sc_mee_generar():
    """Crea las SCs MEE reales en estado 'Pendiente'.

    Acepta sesión 'compras_user' o ?clave=AUTO_PLAN_CRON_KEY (cron Render).
    Body: {modo, origen?, enviar_email?, generico?}
    """
    clave_cron = request.args.get('clave', '') or (request.json or {}).get('clave', '')
    clave_esperada = os.environ.get('AUTO_PLAN_CRON_KEY', '')
    es_cron = bool(clave_esperada and clave_cron == clave_esperada)
    if not es_cron and 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    d = request.json or {}
    modo = (d.get('modo') or request.args.get('modo') or 'mensual').strip()
    origen = (d.get('origen') or request.args.get('origen') or '').strip() or None
    if origen and origen not in ('China', 'Local', 'Mixto'):
        return jsonify({'error': 'origen invalido'}), 400
    enviar_email = d.get('enviar_email', True)
    generico = bool(d.get('generico', False))
    user = 'auto-plan-ia' if es_cron else session.get('compras_user', 'manual')

    conn = get_db(); c = conn.cursor()
    plan = _calcular_auto_sc_mee(conn, modo=modo, origen_filtro=origen, generico=generico)

    scs_creadas = []
    fecha_hoy_iso = datetime.now().date().isoformat()
    for proveedor_clave, items in plan['scs_por_proveedor'].items():
        # extraer proveedor + origen del clave "Proveedor (Origen)"
        prov_real = proveedor_clave.split(' (')[0]
        origen_real = items[0]['origen'] if items else 'Local'

        c.execute("""
            SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)), 0)
            FROM solicitudes_compra WHERE numero LIKE ?
        """, (f"SOL-{datetime.now().strftime('%Y')}-%",))
        n = (c.fetchone()[0] or 0) + 1
        numero = f"SOL-{datetime.now().strftime('%Y')}-{n:04d}"

        observ = (f'🤖 Auto-SC IA MEE ({modo}) · proveedor: {prov_real} · '
                  f'origen: {origen_real} · '
                  f'horizonte {270 if origen_real == "China" else 90}d · '
                  f'buffer IA + estacional + MOQ aplicados')

        c.execute("""
            INSERT INTO solicitudes_compra
              (numero, fecha, estado, solicitante, urgencia, observaciones,
               area, empresa, categoria, tipo, fecha_requerida, valor)
            VALUES (?, ?, 'Pendiente', ?, ?, ?, 'Produccion', 'Espagiria',
                    'Material de Empaque', 'Compra', ?, ?)
        """, (numero, datetime.now().isoformat(), user,
              'Alta' if (modo == 'urgente' or origen_real == 'China') else 'Normal',
              observ, fecha_hoy_iso,
              sum(it.get('valor_estimado', 0) for it in items)))

        for it in items:
            try:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado, proveedor_sugerido)
                    VALUES (?, ?, ?, ?, 'und', ?, ?, ?)
                """, (numero, it['mee_codigo'], it['mee_nombre'],
                      it['cantidad_unidades'], it['justificacion'],
                      it.get('valor_estimado', 0), prov_real))
            except sqlite3.OperationalError:
                # Sin proveedor_sugerido en esquema legacy
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado)
                    VALUES (?, ?, ?, ?, 'und', ?, ?)
                """, (numero, it['mee_codigo'], it['mee_nombre'],
                      it['cantidad_unidades'], it['justificacion'],
                      it.get('valor_estimado', 0)))

        scs_creadas.append({
            'numero': numero,
            'proveedor': prov_real,
            'origen': origen_real,
            'items_count': len(items),
            'total_unidades': round(sum(it['cantidad_unidades'] for it in items), 0),
            'valor_estimado': round(sum(it.get('valor_estimado', 0) for it in items), 2),
        })
    # 2) SC GENÉRICA (Sebastián 1-may-2026): items abiertos para SKUs sin
    # mapping. 1 sola SC con todos los items a asignar por Catalina.
    if generico and plan.get('items_genericos'):
        items_gen = plan['items_genericos']
        c.execute("""
            SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)), 0)
            FROM solicitudes_compra WHERE numero LIKE ?
        """, (f"SOL-{datetime.now().strftime('%Y')}-%",))
        n_gen = (c.fetchone()[0] or 0) + 1
        numero_gen = f"SOL-{datetime.now().strftime('%Y')}-{n_gen:04d}"
        n_skus = len(set(it['sku_codigo'] for it in items_gen))
        observ_gen = (f'🎯 SC GENÉRICA Auto-MEE ({modo}) · {len(items_gen)} items '
                      f'de {n_skus} SKUs SIN mapping · Catalina debe asignar '
                      f'mee_codigo + proveedor en cada item · el sistema '
                      f'aprende y guarda en sku_mee_config para futuras SCs')
        c.execute("""
            INSERT INTO solicitudes_compra
              (numero, fecha, estado, solicitante, urgencia, observaciones,
               area, empresa, categoria, tipo, fecha_requerida, valor)
            VALUES (?, ?, 'Pendiente', ?, 'Normal', ?, 'Produccion', 'Espagiria',
                    'Material de Empaque', 'Compra', ?, 0)
        """, (numero_gen, datetime.now().isoformat(), user, observ_gen,
              fecha_hoy_iso))
        for it in items_gen:
            try:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado, proveedor_sugerido)
                    VALUES (?, '', ?, ?, 'und', ?, 0, '')
                """, (numero_gen,
                      f"[POR-ASIGNAR] {it['descripcion']}",
                      it['cantidad_unidades'], it['justificacion']))
            except sqlite3.OperationalError:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado)
                    VALUES (?, '', ?, ?, 'und', ?, 0)
                """, (numero_gen,
                      f"[POR-ASIGNAR] {it['descripcion']}",
                      it['cantidad_unidades'], it['justificacion']))
        scs_creadas.append({
            'numero': numero_gen,
            'proveedor': '(POR ASIGNAR)',
            'origen': 'generico',
            'items_count': len(items_gen),
            'total_unidades': sum(it['cantidad_unidades'] for it in items_gen),
            'valor_estimado': 0,
        })

    conn.commit()

    # Email a Alejandro
    email_enviado = False
    if enviar_email and scs_creadas:
        try:
            html = _generar_html_auto_sc_mee(plan, scs_creadas, modo)
            destinatarios = []
            try:
                rows = c.execute("""
                    SELECT email FROM email_destinatarios_config
                    WHERE activo=1 AND email != ''
                      AND (rol='gerencia_produccion' OR LOWER(email) LIKE '%alejandro%' OR recibe_compras_aprob=1)
                """).fetchall()
                destinatarios = [r[0] for r in rows if r[0]]
            except Exception:
                pass
            if not destinatarios:
                try:
                    rows = c.execute("SELECT email FROM email_destinatarios_config WHERE activo=1 AND email != ''").fetchall()
                    destinatarios = [r[0] for r in rows if r[0]]
                except Exception:
                    pass
            if destinatarios:
                import threading, sys, os as _os
                sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
                from notificaciones import SistemaNotificaciones
                notif = SistemaNotificaciones()
                threading.Thread(
                    target=notif._enviar_email,
                    args=(f'🤖 Auto-SC IA MEE · {len(scs_creadas)} SCs creadas ({modo})', html, destinatarios),
                    daemon=True
                ).start()
                email_enviado = True
        except Exception as e:
            log.warning(f'Email auto-SC MEE fallo: {e}')

    # Log
    try:
        notas = (f'{plan["kpis"]["total_items"]} items · '
                 f'{len(scs_creadas)} SCs · email={email_enviado} · '
                 f'origen={origen or "todos"}')
        c.execute("""
            INSERT INTO auto_plan_runs
              (ejecutado_at, ejecutado_por, tipo, horizonte_dias,
               producciones_creadas, compras_creadas, alertas_criticas,
               emails_enviados, error, payload_json)
            VALUES (?, ?, ?, ?, 0, ?, 0, ?, NULL, ?)
        """, (datetime.now().isoformat(), user,
              f'auto_sc_mee_{modo}',
              270 if (origen == 'China') else (90 if modo == 'mensual' else 30),
              len(scs_creadas),
              1 if email_enviado else 0,
              json.dumps({'modo': modo, 'origen': origen,
                          'kpis': plan['kpis'], 'scs': scs_creadas,
                          'notas': notas}, default=str)))
        conn.commit()
    except Exception as e:
        log.warning(f'Log auto-sc-mee fallo: {e}')

    return jsonify({
        'ok': True,
        'modo': modo,
        'origen': origen,
        'scs_creadas': scs_creadas,
        'items_huerfanos': plan['items_huerfanos'],
        'email_enviado': email_enviado,
        'kpis': plan['kpis'],
        'razon_vacio': plan.get('razon_vacio'),
        'mensaje': (f'✅ {len(scs_creadas)} SCs MEE creadas · '
                    f'{plan["kpis"]["total_items"]} items · '
                    f'email={"sí" if email_enviado else "no"}'),
    })


def _generar_html_auto_sc_mee(plan, scs_creadas, modo):
    """HTML email para Alejandro con resumen de SCs MEE creadas."""
    fecha_hoy = datetime.now().date()
    titulo = '🚨 SC MEE Urgentes' if modo == 'urgente' else '🤖 SC MEE Mensuales (China 9m + Local 90d)'
    bgHeader = '#dc2626' if modo == 'urgente' else '#0f766e'

    html_scs = ''
    for sc in scs_creadas:
        items = plan['scs_por_proveedor'].get(f"{sc['proveedor']} ({sc['origen']})", [])
        flag_origen = '🇨🇳' if sc['origen'] == 'China' else '🇨🇴'
        html_scs += f'<div style="background:#f8fafc;border-left:4px solid {bgHeader};padding:10px 14px;margin-bottom:8px;border-radius:0 6px 6px 0">'
        html_scs += f'<b>{sc["numero"]}</b> · {flag_origen} {sc["proveedor"]} · {sc["items_count"]} MEE · {int(sc["total_unidades"]):,} ud'
        if sc.get('valor_estimado'):
            html_scs += f' · ${sc["valor_estimado"]:,.0f}'
        html_scs += '<ul style="margin:6px 0 0 18px;padding:0;font-size:11px;color:#475569">'
        for it in items[:6]:
            html_scs += f'<li><b>{it["mee_nombre"]}</b>: {int(it["cantidad_unidades"]):,} ud (lead {it["lead_time_dias"]}d)</li>'
        if len(items) > 6:
            html_scs += f'<li style="color:#94a3b8">+ {len(items)-6} MEE más</li>'
        html_scs += '</ul></div>'

    huerfanos_html = ''
    if plan['items_huerfanos']:
        huerfanos_html = '<div style="background:#fef3c7;border:1px solid #fcd34d;padding:10px;border-radius:6px;margin-top:14px">'
        huerfanos_html += f'<b style="color:#92400e">⚠️ {len(plan["items_huerfanos"])} MEE SIN proveedor</b>'
        huerfanos_html += '<div style="font-size:11px;color:#78350f;margin-top:4px">No se crearon SCs (mee_lead_time_config sin proveedor_principal):</div>'
        huerfanos_html += '<ul style="margin:4px 0 0 18px;padding:0;font-size:11px;color:#78350f">'
        for it in plan['items_huerfanos'][:8]:
            huerfanos_html += f'<li>{it["mee_nombre"]}: {int(it["cantidad_unidades"]):,} ud ({it["origen"]})</li>'
        huerfanos_html += '</ul></div>'

    return f'''<!DOCTYPE html>
<html><body style="font-family:-apple-system,sans-serif;background:#f3f4f6;padding:20px;margin:0">
  <div style="max-width:640px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,.08)">
    <div style="background:{bgHeader};color:#fff;padding:20px">
      <h2 style="margin:0;font-size:20px">{titulo}</h2>
      <p style="margin:4px 0 0;opacity:.9;font-size:13px">{fecha_hoy.strftime("%d-%b-%Y")} · Auto-Plan IA · MEE</p>
    </div>
    <div style="padding:20px">
      <div style="background:#ecfdf5;border:1px solid #6ee7b7;padding:10px;border-radius:6px;margin-bottom:14px;font-size:13px;color:#065f46">
        ✅ <b>{len(scs_creadas)} solicitudes creadas</b> · {plan["kpis"]["total_items"]} MEE · {plan["kpis"]["total_unidades"]:,.0f} unidades<br>
        Estado: <b>Pendiente</b> · Catalina y Alejandro revisan en <a href="/solicitudes" style="color:#065f46">/solicitudes</a>
      </div>
      <h3 style="color:#0f172a;font-size:14px;margin:14px 0 8px">📦 Solicitudes MEE creadas</h3>
      {html_scs}
      {huerfanos_html}
      <div style="margin-top:18px;padding:12px;background:#f1f5f9;border-radius:8px;font-size:11px;color:#64748b">
        Horizontes: 🇨🇳 China 9m (270d) · 🇨🇴 Local 90d · Urgente 30d.<br>
        MOQ aplicado por proveedor. Buffer IA tendencia + estacional aplicados.<br>
        Serigrafía/tampografía van por cron D-20 (no incluidas aquí). Etiquetas se piden post-envasado.
      </div>
    </div>
  </div>
</body></html>'''


# ════════════════════════════════════════════════════════════════════════
# ALERTAS MEE · etiquetas post-envasado + serigrafía/tampografía D-20
# ════════════════════════════════════════════════════════════════════════
# Sebastián (1-may-2026): etiquetas se piden al envasar (no proyectivo).
# Serigrafía/tampografía 20d antes (D-20). Esta capa no PIDE solo, sino
# que ALERTA. Catalina/Alejandro deciden con un click.

@bp.route('/api/planta/alerta-etiquetas-pendientes', methods=['GET'])
def alerta_etiquetas_pendientes():
    """Lista envasados recientes para que Catalina valide si ya pidió etiquetas.

    Query: ?dias=14 (default 14)
    Returns: lista de envasados + flag si ya hay SC etiqueta asociada.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    dias = int(request.args.get('dias', 14))
    conn = get_db(); c = conn.cursor()

    # Envasados últimos N días
    envasados = []
    try:
        rows = c.execute("""
            SELECT id, lote, producto, presentacion, unidades, fecha,
                   COALESCE(envase_codigo,''), COALESCE(tapa_codigo,'')
            FROM envasado
            WHERE COALESCE(estado,'Completado')='Completado'
              AND date(fecha) >= date('now','-' || ? || ' days')
            ORDER BY fecha DESC, id DESC
        """, (dias,)).fetchall()
        envasados = [dict(zip(['id','lote','producto','presentacion','unidades',
                                'fecha','envase_codigo','tapa_codigo'], r)) for r in rows]
    except Exception:
        pass

    # Para cada envasado, ¿hay SC de etiqueta creada después?
    # Heurística: SC con categoria='Material de Empaque' Y solicitudes_compra_items
    # contiene el producto en justificación o nombre_mp tiene 'etiqueta'.
    for env in envasados:
        try:
            tiene_sc = c.execute("""
                SELECT COUNT(DISTINCT s.id)
                FROM solicitudes_compra s
                  JOIN solicitudes_compra_items i ON i.numero = s.numero
                WHERE s.categoria='Material de Empaque'
                  AND date(s.fecha) >= ?
                  AND (LOWER(i.nombre_mp) LIKE '%etiqueta%'
                       OR LOWER(s.observaciones) LIKE '%etiqueta%')
                  AND (LOWER(s.observaciones) LIKE ? OR LOWER(i.justificacion) LIKE ?)
            """, (env['fecha'], f"%{env['lote'].lower()}%",
                  f"%{env['lote'].lower()}%")).fetchone()[0]
            env['tiene_sc_etiqueta'] = bool(tiene_sc)
        except Exception:
            env['tiene_sc_etiqueta'] = False

        # Etiquetas asignadas al SKU/producto en sku_mee_config
        try:
            etqs = c.execute("""
                SELECT DISTINCT s.mee_codigo, m.descripcion, m.stock_actual,
                       cfg.proveedor_principal
                FROM sku_mee_config s
                  LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
                  LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = s.mee_codigo
                WHERE s.aplica = 1
                  AND s.componente_tipo = 'etiqueta'
                  AND UPPER(TRIM(s.sku_codigo)) = UPPER(TRIM(?))
            """, (env['producto'],)).fetchall()
            env['etiquetas_sku'] = [{'codigo': r[0], 'nombre': r[1] or r[0],
                                      'stock': r[2] or 0,
                                      'proveedor': r[3] or ''} for r in etqs]
        except Exception:
            env['etiquetas_sku'] = []

    return jsonify({
        'dias_horizonte': dias,
        'envasados': envasados,
        'pendientes': [e for e in envasados if not e['tiene_sc_etiqueta']],
        'kpis': {
            'total_envasados': len(envasados),
            'pendientes_etiqueta': sum(1 for e in envasados if not e['tiene_sc_etiqueta']),
            'unidades_pendientes': sum(e['unidades'] or 0 for e in envasados if not e['tiene_sc_etiqueta']),
        },
    })


@bp.route('/api/planta/alerta-d20-pendientes', methods=['GET'])
def alerta_d20_pendientes():
    """Lista producciones próximas (D-15 a D-25) que necesitan serigrafía/tampografía.

    Sebastian: "serigrafia ideal pedir 20 dias antes de la producción al
    igual que tampografia". Esta alerta lista producciones futuras que caen
    en ventana D-20 ± 5d, con sus componentes de decoración asociados.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()
    d_min = fecha_hoy + timedelta(days=15)
    d_max = fecha_hoy + timedelta(days=25)

    # Eventos del Calendar en ventana
    eventos = _calendar_events_cached()
    skus_activos = {}
    try:
        rows = c.execute("""
            SELECT producto_nombre FROM sku_planeacion_config
            WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
        """).fetchall()
        skus_activos = {r[0]: r[0] for r in rows}
    except Exception:
        pass

    producciones_d20 = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if f < d_min or f > d_max:
            continue
        # Match con SKU activo
        producto_match = None
        for prod_nom in skus_activos.keys():
            try:
                alias = _alias_calendar_for(c, prod_nom)
                score = _match_producto_evento(prod_nom, alias, ev.get('titulo'),
                                                ev.get('descripcion', ''))
                if score >= 60:
                    producto_match = prod_nom
                    break
            except Exception:
                continue
        if not producto_match:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion','')) or 30
        dias_hasta = (f - fecha_hoy).days

        # Buscar serigrafía/tampografía asociada al SKU/producto
        decoraciones = []
        try:
            rows = c.execute("""
                SELECT DISTINCT s.mee_codigo, s.componente_tipo, m.descripcion,
                       m.stock_actual, cfg.proveedor_principal, cfg.lead_time_dias
                FROM sku_mee_config s
                  LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
                  LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = s.mee_codigo
                WHERE s.aplica = 1
                  AND s.componente_tipo IN ('serigrafia','tampografia')
                  AND COALESCE(cfg.disparo_d20, 0) = 1
                  AND UPPER(TRIM(s.sku_codigo)) = UPPER(TRIM(?))
            """, (producto_match,)).fetchall()
            decoraciones = [{
                'codigo': r[0], 'tipo': r[1], 'nombre': r[2] or r[0],
                'stock': r[3] or 0, 'proveedor': r[4] or '',
                'lead_time': r[5] or 20,
            } for r in rows]
        except Exception:
            pass

        producciones_d20.append({
            'fecha': f.isoformat(),
            'dias_hasta': dias_hasta,
            'titulo': ev.get('titulo', ''),
            'producto': producto_match,
            'kg': kg,
            'unidades_estimadas': int(kg * 1000 / 30),  # asume 30g/SKU promedio
            'decoraciones': decoraciones,
            'critico': dias_hasta <= 18 and bool(decoraciones),
        })

    # Sort por urgencia (más cercanas primero)
    producciones_d20.sort(key=lambda p: p['dias_hasta'])

    return jsonify({
        'fecha_hoy': fecha_hoy.isoformat(),
        'ventana': {'desde': d_min.isoformat(), 'hasta': d_max.isoformat()},
        'producciones': producciones_d20,
        'kpis': {
            'total': len(producciones_d20),
            'criticas': sum(1 for p in producciones_d20 if p['critico']),
            'sin_decoraciones': sum(1 for p in producciones_d20 if not p['decoraciones']),
        },
    })


@bp.route('/api/planta/sc-etiqueta-rapida', methods=['POST'])
def sc_etiqueta_rapida():
    """Crea SC rápida de etiquetas para un envasado específico.

    Body: {envasado_id, codigos_etiqueta?: [str]}
    Si codigos_etiqueta no se pasa, toma todas las etiquetas mapeadas al SKU.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    env_id = d.get('envasado_id')
    if not env_id:
        return jsonify({'error': 'envasado_id requerido'}), 400

    conn = get_db(); c = conn.cursor()
    env = c.execute("""
        SELECT lote, producto, presentacion, unidades, fecha
        FROM envasado WHERE id = ?
    """, (env_id,)).fetchone()
    if not env:
        return jsonify({'error': 'envasado no encontrado'}), 404
    lote, producto, presentacion, unidades, fecha = env

    # Etiquetas del SKU
    rows = c.execute("""
        SELECT s.mee_codigo, s.cantidad_por_unidad, m.descripcion,
               cfg.proveedor_principal, cfg.precio_unit
        FROM sku_mee_config s
          LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
          LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = s.mee_codigo
        WHERE s.aplica = 1
          AND s.componente_tipo = 'etiqueta'
          AND UPPER(TRIM(s.sku_codigo)) = UPPER(TRIM(?))
    """, (producto,)).fetchall()
    if not rows:
        return jsonify({'error': f'No hay etiquetas configuradas para SKU {producto}'}), 400

    codigos_filtro = d.get('codigos_etiqueta') or []
    items = []
    for cod, cant_pu, desc, prov, prec in rows:
        if codigos_filtro and cod not in codigos_filtro:
            continue
        cant = (unidades or 0) * float(cant_pu or 1)
        items.append({
            'mee_codigo': cod, 'nombre': desc or cod,
            'cantidad': cant, 'proveedor': prov or '',
            'precio_unit': float(prec or 0),
            'valor_estimado': cant * float(prec or 0),
        })
    if not items:
        return jsonify({'error': 'Sin items para crear'}), 400

    # 1 SC por proveedor (igual que el Auto-SC general)
    scs_creadas = []
    items_huerfanos = []
    items_por_prov = {}
    for it in items:
        prov = it['proveedor'] or '(sin proveedor)'
        if prov == '(sin proveedor)':
            items_huerfanos.append(it)
            continue
        items_por_prov.setdefault(prov, []).append(it)

    fecha_hoy_iso = datetime.now().date().isoformat()
    for prov, prov_items in items_por_prov.items():
        n = c.execute("""
            SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)), 0)
            FROM solicitudes_compra WHERE numero LIKE ?
        """, (f"SOL-{datetime.now().strftime('%Y')}-%",)).fetchone()[0] + 1
        numero = f"SOL-{datetime.now().strftime('%Y')}-{n:04d}"
        observ = (f'🏷️ SC etiqueta post-envasado · lote {lote} · {producto} '
                  f'· {presentacion} · {unidades} ud · proveedor {prov}')
        c.execute("""
            INSERT INTO solicitudes_compra
              (numero, fecha, estado, solicitante, urgencia, observaciones,
               area, empresa, categoria, tipo, fecha_requerida, valor)
            VALUES (?, ?, 'Pendiente', ?, 'Alta', ?, 'Produccion', 'Espagiria',
                    'Material de Empaque', 'Compra', ?, ?)
        """, (numero, datetime.now().isoformat(), user, observ, fecha_hoy_iso,
              sum(it['valor_estimado'] for it in prov_items)))
        for it in prov_items:
            try:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado, proveedor_sugerido)
                    VALUES (?, ?, ?, ?, 'und', ?, ?, ?)
                """, (numero, it['mee_codigo'], it['nombre'], it['cantidad'],
                      f"Etiqueta para envasado lote {lote} · {it['cantidad']:.0f} ud",
                      it['valor_estimado'], it['proveedor']))
            except sqlite3.OperationalError:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado)
                    VALUES (?, ?, ?, ?, 'und', ?, ?)
                """, (numero, it['mee_codigo'], it['nombre'], it['cantidad'],
                      f"Etiqueta para envasado lote {lote}", it['valor_estimado']))
        scs_creadas.append({'numero': numero, 'proveedor': prov,
                             'items': len(prov_items),
                             'total_valor': sum(it['valor_estimado'] for it in prov_items)})

    conn.commit()
    return jsonify({
        'ok': True,
        'envasado_id': env_id,
        'lote': lote,
        'producto': producto,
        'scs_creadas': scs_creadas,
        'items_huerfanos': items_huerfanos,
        'mensaje': f'✅ {len(scs_creadas)} SCs etiqueta creadas',
    })


@bp.route('/api/planta/auto-d20-cron', methods=['POST'])
def auto_d20_cron():
    """Cron diario: revisa producciones futuras D-15..D-25 y crea SCs
    automáticas para serigrafía/tampografía si aún no existen.

    Sebastián (1-may-2026): "cron diario automatico" para D-20.

    Acepta sesión 'compras_user' o ?clave=AUTO_PLAN_CRON_KEY (cron Render).
    Body opcional: {dry_run: bool, dias_ventana: 5}
    """
    clave_cron = request.args.get('clave', '') or (request.json or {}).get('clave', '')
    clave_esperada = os.environ.get('AUTO_PLAN_CRON_KEY', '')
    es_cron = bool(clave_esperada and clave_cron == clave_esperada)
    if not es_cron and 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = 'auto-plan-ia' if es_cron else session.get('compras_user', 'manual')
    d = request.json or {}
    dry_run = bool(d.get('dry_run', False))

    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()

    # Llamar internamente a la lógica de alerta-d20
    eventos = _calendar_events_cached()
    d_min = fecha_hoy + timedelta(days=15)
    d_max = fecha_hoy + timedelta(days=25)

    skus_activos = {}
    try:
        rows = c.execute("""
            SELECT producto_nombre FROM sku_planeacion_config
            WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
        """).fetchall()
        skus_activos = {r[0]: r[0] for r in rows}
    except Exception:
        pass

    candidatos = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if f < d_min or f > d_max:
            continue
        producto_match = None
        for prod_nom in skus_activos.keys():
            try:
                alias = _alias_calendar_for(c, prod_nom)
                score = _match_producto_evento(prod_nom, alias, ev.get('titulo'),
                                                ev.get('descripcion', ''))
                if score >= 60:
                    producto_match = prod_nom
                    break
            except Exception:
                continue
        if not producto_match:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion','')) or 30
        unidades_estimadas = int(kg * 1000 / 30)
        # Verificar si ya existe SC d20 para esta fecha+producto
        existe = c.execute("""
            SELECT 1 FROM solicitudes_compra
            WHERE categoria='Servicios'
              AND observaciones LIKE ?
              AND date(fecha) >= date('now','-30 days')
            LIMIT 1
        """, (f'%decoración D-20 · {producto_match}%',)).fetchone()
        if existe:
            continue
        candidatos.append({
            'producto': producto_match,
            'fecha': f.isoformat(),
            'unidades_estimadas': unidades_estimadas,
            'kg': kg,
        })

    scs_creadas = []
    for cand in candidatos:
        # Buscar componentes serigrafía/tampografía mapeados al SKU
        rows = c.execute("""
            SELECT s.mee_codigo, s.componente_tipo, s.cantidad_por_unidad,
                   m.descripcion, cfg.proveedor_principal, cfg.precio_unit
            FROM sku_mee_config s
              LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
              LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = s.mee_codigo
            WHERE s.aplica = 1
              AND s.componente_tipo IN ('serigrafia','tampografia')
              AND COALESCE(cfg.disparo_d20,0) = 1
              AND UPPER(TRIM(s.sku_codigo)) = UPPER(TRIM(?))
        """, (cand['producto'],)).fetchall()
        if not rows:
            continue

        items_por_prov = {}
        for cod, tipo, cant_pu, desc, prov, prec in rows:
            cant = cand['unidades_estimadas'] * float(cant_pu or 1)
            it = {'mee_codigo': cod, 'tipo': tipo, 'nombre': desc or cod,
                  'cantidad': cant, 'proveedor': prov or '',
                  'precio_unit': float(prec or 0),
                  'valor_estimado': cant * float(prec or 0)}
            if it['proveedor']:
                items_por_prov.setdefault(prov, []).append(it)

        for prov, prov_items in items_por_prov.items():
            if dry_run:
                scs_creadas.append({
                    'producto': cand['producto'],
                    'fecha': cand['fecha'],
                    'proveedor': prov,
                    'items': len(prov_items),
                    'unidades': cand['unidades_estimadas'],
                    'dry_run': True,
                })
                continue
            n = c.execute("""
                SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)), 0)
                FROM solicitudes_compra WHERE numero LIKE ?
            """, (f"SOL-{datetime.now().strftime('%Y')}-%",)).fetchone()[0] + 1
            numero = f"SOL-{datetime.now().strftime('%Y')}-{n:04d}"
            observ = (f'🎨 Cron D-20 · decoración D-20 · {cand["producto"]} · '
                      f'producción {cand["fecha"]} · {cand["unidades_estimadas"]} ud · '
                      f'proveedor {prov}')
            c.execute("""
                INSERT INTO solicitudes_compra
                  (numero, fecha, estado, solicitante, urgencia, observaciones,
                   area, empresa, categoria, tipo, fecha_requerida, valor)
                VALUES (?, ?, 'Pendiente', ?, 'Alta', ?, 'Produccion', 'Espagiria',
                        'Servicios', 'Compra', ?, ?)
            """, (numero, datetime.now().isoformat(), user, observ, cand['fecha'],
                  sum(it['valor_estimado'] for it in prov_items)))
            for it in prov_items:
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado, proveedor_sugerido)
                        VALUES (?, ?, ?, ?, 'und', ?, ?, ?)
                    """, (numero, it['mee_codigo'], it['nombre'], it['cantidad'],
                          f"{it['tipo']} D-20 cron · {cand['producto']} · {cand['unidades_estimadas']} ud",
                          it['valor_estimado'], it['proveedor']))
                except sqlite3.OperationalError:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado)
                        VALUES (?, ?, ?, ?, 'und', ?, ?)
                    """, (numero, it['mee_codigo'], it['nombre'], it['cantidad'],
                          f"{it['tipo']} D-20 cron · {cand['producto']}",
                          it['valor_estimado']))
            scs_creadas.append({
                'numero': numero,
                'producto': cand['producto'],
                'fecha': cand['fecha'],
                'proveedor': prov,
                'items': len(prov_items),
                'unidades': cand['unidades_estimadas'],
            })

    if not dry_run:
        conn.commit()
        # Log
        try:
            c.execute("""
                INSERT INTO auto_plan_runs
                  (ejecutado_at, ejecutado_por, tipo, horizonte_dias,
                   producciones_creadas, compras_creadas, alertas_criticas,
                   emails_enviados, error, payload_json)
                VALUES (?, ?, 'auto_d20_cron', 20, 0, ?, 0, 0, NULL, ?)
            """, (datetime.now().isoformat(), user, len(scs_creadas),
                  json.dumps({'scs': scs_creadas, 'candidatos': len(candidatos)},
                              default=str)))
            conn.commit()
        except Exception:
            pass

    return jsonify({
        'ok': True,
        'dry_run': dry_run,
        'fecha_hoy': fecha_hoy.isoformat(),
        'producciones_en_ventana': len(candidatos),
        'scs_creadas': scs_creadas,
        'mensaje': (f'{"DRY RUN: " if dry_run else ""}'
                    f'{len(scs_creadas)} SCs decoración '
                    f'{"a crear" if dry_run else "creadas"} '
                    f'para {len(set(s.get("producto") for s in scs_creadas))} producciones'),
    })


@bp.route('/api/planta/sc-d20-rapida', methods=['POST'])
def sc_d20_rapida():
    """Crea SC rápida de serigrafía/tampografía para una producción D-20.

    Body: {producto, fecha_produccion, unidades_estimadas}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    producto = (d.get('producto') or '').strip()
    fecha_prod = (d.get('fecha_produccion') or '').strip()
    unidades = int(d.get('unidades_estimadas') or 0)
    if not producto or not fecha_prod or unidades <= 0:
        return jsonify({'error': 'producto, fecha_produccion, unidades_estimadas requeridos'}), 400

    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT s.mee_codigo, s.componente_tipo, s.cantidad_por_unidad,
               m.descripcion, cfg.proveedor_principal, cfg.precio_unit
        FROM sku_mee_config s
          LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
          LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = s.mee_codigo
        WHERE s.aplica = 1
          AND s.componente_tipo IN ('serigrafia','tampografia')
          AND COALESCE(cfg.disparo_d20,0) = 1
          AND UPPER(TRIM(s.sku_codigo)) = UPPER(TRIM(?))
    """, (producto,)).fetchall()
    if not rows:
        return jsonify({'error': f'No hay serigrafía/tampografía D-20 configurada para {producto}'}), 400

    items_por_prov = {}
    items_huerfanos = []
    for cod, tipo, cant_pu, desc, prov, prec in rows:
        cant = unidades * float(cant_pu or 1)
        item = {
            'mee_codigo': cod, 'tipo': tipo, 'nombre': desc or cod,
            'cantidad': cant, 'proveedor': prov or '',
            'precio_unit': float(prec or 0),
            'valor_estimado': cant * float(prec or 0),
        }
        if not item['proveedor']:
            items_huerfanos.append(item)
            continue
        items_por_prov.setdefault(prov, []).append(item)

    scs_creadas = []
    fecha_hoy_iso = datetime.now().date().isoformat()
    for prov, prov_items in items_por_prov.items():
        n = c.execute("""
            SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)), 0)
            FROM solicitudes_compra WHERE numero LIKE ?
        """, (f"SOL-{datetime.now().strftime('%Y')}-%",)).fetchone()[0] + 1
        numero = f"SOL-{datetime.now().strftime('%Y')}-{n:04d}"
        observ = (f'🎨 SC decoración D-20 · {producto} · producción {fecha_prod} '
                  f'· {unidades} ud · proveedor {prov}')
        c.execute("""
            INSERT INTO solicitudes_compra
              (numero, fecha, estado, solicitante, urgencia, observaciones,
               area, empresa, categoria, tipo, fecha_requerida, valor)
            VALUES (?, ?, 'Pendiente', ?, 'Alta', ?, 'Produccion', 'Espagiria',
                    'Servicios', 'Compra', ?, ?)
        """, (numero, datetime.now().isoformat(), user, observ, fecha_prod,
              sum(it['valor_estimado'] for it in prov_items)))
        for it in prov_items:
            try:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado, proveedor_sugerido)
                    VALUES (?, ?, ?, ?, 'und', ?, ?, ?)
                """, (numero, it['mee_codigo'], it['nombre'], it['cantidad'],
                      f"{it['tipo']} para producción {fecha_prod} · {producto} · {unidades} ud",
                      it['valor_estimado'], it['proveedor']))
            except sqlite3.OperationalError:
                c.execute("""
                    INSERT INTO solicitudes_compra_items
                      (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       justificacion, valor_estimado)
                    VALUES (?, ?, ?, ?, 'und', ?, ?)
                """, (numero, it['mee_codigo'], it['nombre'], it['cantidad'],
                      f"{it['tipo']} para producción {fecha_prod}",
                      it['valor_estimado']))
        scs_creadas.append({'numero': numero, 'proveedor': prov,
                             'items': len(prov_items),
                             'total_valor': sum(it['valor_estimado'] for it in prov_items)})

    conn.commit()
    return jsonify({
        'ok': True,
        'producto': producto,
        'fecha_produccion': fecha_prod,
        'unidades': unidades,
        'scs_creadas': scs_creadas,
        'items_huerfanos': items_huerfanos,
        'mensaje': f'✅ {len(scs_creadas)} SCs decoración creadas',
    })


# ════════════════════════════════════════════════════════════════════════
# NORMALIZACIÓN MEE · backfill proveedor + auto-mapping SKU→MEE fuzzy
# ════════════════════════════════════════════════════════════════════════
# Sebastián (1-may-2026): "no esta tomando nada para eso podemos normalizar".
# 82 MEE configurados pero 0 con proveedor + 1 SKU mapeado de 32 activos.
# Datos reales viven en maestro_mee.proveedor (no se copiaron al config).
# Más: muchas etiquetas/serigrafías tienen el nombre del SKU en su descripción
# → auto-mapping fuzzy posible.

def _normalizar_palabras(s):
    """Tokeniza un string para comparación fuzzy: minúsculas, sin acentos,
    quita conectores, devuelve set de palabras significativas."""
    if not s:
        return set()
    import unicodedata
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').lower()
    # Quitar puntuación
    import re
    s = re.sub(r'[^a-z0-9\s+]', ' ', s)
    palabras = set(s.split())
    # Quitar conectores comunes
    stopwords = {'de', 'del', 'la', 'el', 'y', 'a', 'en', 'con', 'sin', 'para',
                 'por', 'al', 'los', 'las', 'un', 'una', 'es', 'que', 'lo'}
    return {p for p in palabras if p not in stopwords and len(p) >= 2}


def _score_fuzzy(palabras_a, palabras_b):
    """Score 0-100 de similitud entre dos sets de palabras (Jaccard expandido)."""
    if not palabras_a or not palabras_b:
        return 0
    inter = palabras_a & palabras_b
    if not inter:
        return 0
    # Pesar más palabras "raras" (no comunes en muchos MEE)
    score = (len(inter) / max(len(palabras_a), len(palabras_b))) * 100
    # Bonus si todas las palabras del SKU están en el MEE
    if palabras_a.issubset(palabras_b):
        score = min(100, score + 25)
    return round(score, 1)


# ════════════════════════════════════════════════════════════════════════
# APRENDIZAJE MEE · Catalina asigna código MEE a item genérico
# ════════════════════════════════════════════════════════════════════════
# Sebastián (1-may-2026): "sea catalina quien asigne que se hace en compras,
# serigrafia, tampografia etiqueta, proveerdor y a cual corresponde y cuando
# se use se descuenta automatico? y asi se van guardando para el futuro
# porque hoy en dia no sabemos a quien corresponde cada cosa".

@bp.route('/api/planta/items-por-asignar', methods=['GET'])
def items_por_asignar():
    """Lista items de SCs MEE genéricas con codigo_mp vacío + nombre POR-ASIGNAR.
    Para que Catalina los asigne desde el panel Auto-SC MEE.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT i.id, i.numero, i.codigo_mp, i.nombre_mp, i.cantidad_g,
               i.unidad, i.justificacion, s.fecha, s.estado, s.solicitante
        FROM solicitudes_compra_items i
          JOIN solicitudes_compra s ON s.numero = i.numero
        WHERE COALESCE(i.codigo_mp, '') = ''
          AND (i.nombre_mp LIKE '%POR-ASIGNAR%' OR i.nombre_mp LIKE '%POR ASIGNAR%')
          AND s.categoria = 'Material de Empaque'
          AND s.estado IN ('Pendiente', 'En Revision')
        ORDER BY s.fecha DESC, i.id ASC
    """).fetchall()
    cols = ['id','numero','codigo_mp','nombre_mp','cantidad_g','unidad',
            'justificacion','fecha_sc','estado_sc','solicitante']
    items = [dict(zip(cols, r)) for r in rows]

    # Catálogo MEE para dropdown (codigo + descripcion + categoria)
    mees = []
    try:
        for r in c.execute("""
            SELECT codigo, descripcion, categoria, stock_actual
            FROM maestro_mee
            WHERE COALESCE(estado,'Activo')='Activo'
            ORDER BY categoria, codigo
        """).fetchall():
            mees.append({'codigo': r[0], 'descripcion': r[1] or r[0],
                          'categoria': r[2] or '', 'stock': r[3] or 0})
    except Exception:
        pass

    # Catálogo proveedores
    proveedores = []
    try:
        for r in c.execute("SELECT DISTINCT nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall():
            proveedores.append(r[0])
    except Exception:
        try:
            for r in c.execute("SELECT DISTINCT proveedor FROM maestro_mee WHERE COALESCE(proveedor,'') != '' ORDER BY proveedor").fetchall():
                proveedores.append(r[0])
        except Exception:
            pass

    return jsonify({
        'items': items,
        'total': len(items),
        'maestro_mee': mees,
        'proveedores': proveedores,
    })


@bp.route('/api/planta/sc-mee-asignar', methods=['POST'])
def sc_mee_asignar():
    """Catalina asigna un código MEE específico a un item de SC genérica
    y el sistema APRENDE: guarda el mapping en sku_mee_config para próximas
    SCs (no vuelve a generar genérico para ese SKU+componente).

    Body:
      sc_item_id: int — id del solicitudes_compra_items
      mee_codigo: str — código MEE asignado (de maestro_mee)
      proveedor: str (opcional) — actualizar también mee_lead_time_config
      cantidad_por_unidad: float (default 1)

    Acciones:
      1. UPDATE solicitudes_compra_items: codigo_mp = mee_codigo,
         nombre_mp = descripción real, proveedor_sugerido
      2. INSERT OR IGNORE sku_mee_config: aprende para futuro
      3. Si proveedor pasado, UPDATE mee_lead_time_config.proveedor_principal
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    sc_item_id = d.get('sc_item_id')
    mee_codigo = (d.get('mee_codigo') or '').strip()
    proveedor = (d.get('proveedor') or '').strip()
    cantidad_pu = float(d.get('cantidad_por_unidad', 1))
    if not sc_item_id or not mee_codigo:
        return jsonify({'error': 'sc_item_id y mee_codigo requeridos'}), 400

    conn = get_db(); c = conn.cursor()

    # Validar mee_codigo existe
    mee = c.execute(
        "SELECT codigo, descripcion, categoria FROM maestro_mee WHERE codigo = ?",
        (mee_codigo,)
    ).fetchone()
    if not mee:
        return jsonify({'error': f'MEE {mee_codigo} no existe'}), 400

    # Recuperar item original (para extraer SKU + componente del nombre_mp)
    item = c.execute("""
        SELECT id, numero, codigo_mp, nombre_mp, justificacion
        FROM solicitudes_compra_items WHERE id = ?
    """, (sc_item_id,)).fetchone()
    if not item:
        return jsonify({'error': 'Item de SC no encontrado'}), 404
    item_id, sc_numero, codigo_actual, nombre, justif = item

    # Extraer SKU + tipo componente del nombre/justificación
    # Formato esperado: "[POR-ASIGNAR] 📦 Envase primario para SUERO HIDRATANTE AH 1.5%"
    sku = ''
    componente_tipo = 'envase'  # default
    import re
    m = re.search(r'para\s+([^·]+?)(?:\s*·|$)', nombre or '', re.IGNORECASE)
    if m:
        sku = m.group(1).strip()
    if 'envase' in (nombre or '').lower():
        componente_tipo = 'envase'
    elif 'tapa' in (nombre or '').lower():
        componente_tipo = 'tapa'
    elif 'etiqueta' in (nombre or '').lower():
        componente_tipo = 'etiqueta'
    elif 'serigraf' in (nombre or '').lower():
        componente_tipo = 'serigrafia'
    elif 'tampograf' in (nombre or '').lower():
        componente_tipo = 'tampografia'
    elif 'caja' in (nombre or '').lower():
        componente_tipo = 'caja'
    # Permitir override explícito desde body
    if d.get('sku_codigo'):
        sku = d['sku_codigo'].strip()
    if d.get('componente_tipo'):
        componente_tipo = d['componente_tipo'].strip()

    if not sku:
        return jsonify({'error': 'No pude inferir SKU; pasa sku_codigo explícito en body'}), 400

    # 1) UPDATE item de la SC
    nuevo_nombre = mee[1] or mee_codigo
    try:
        c.execute("""
            UPDATE solicitudes_compra_items
               SET codigo_mp = ?, nombre_mp = ?, proveedor_sugerido = ?,
                   actualizado_at = datetime('now'), actualizado_por = ?
             WHERE id = ?
        """, (mee_codigo, nuevo_nombre, proveedor or '', user, item_id))
    except sqlite3.OperationalError:
        c.execute("""
            UPDATE solicitudes_compra_items
               SET codigo_mp = ?, nombre_mp = ?
             WHERE id = ?
        """, (mee_codigo, nuevo_nombre, item_id))

    # 2) APRENDER: guardar en sku_mee_config (idempotente)
    aprendizaje = False
    try:
        c.execute("""
            INSERT OR IGNORE INTO sku_mee_config
              (sku_codigo, mee_codigo, componente_tipo, cantidad_por_unidad,
               aplica, notas)
            VALUES (?, ?, ?, ?, 1, ?)
        """, (sku, mee_codigo, componente_tipo, cantidad_pu,
              f'Aprendido de SC {sc_numero} item #{item_id} por {user}'))
        if c.rowcount > 0:
            aprendizaje = True
    except sqlite3.IntegrityError:
        pass  # ya existe el mapping

    # 3) Si pasaron proveedor, actualizar mee_lead_time_config (upsert)
    if proveedor:
        existing = c.execute(
            "SELECT 1 FROM mee_lead_time_config WHERE mee_codigo = ?", (mee_codigo,)
        ).fetchone()
        if existing:
            c.execute("""
                UPDATE mee_lead_time_config
                   SET proveedor_principal = ?, actualizado_en = datetime('now'),
                       actualizado_por = ?
                 WHERE mee_codigo = ?
            """, (proveedor, user, mee_codigo))
        else:
            c.execute("""
                INSERT INTO mee_lead_time_config
                  (mee_codigo, proveedor_principal, origen, lead_time_dias,
                   moq_unidades, precio_unit, aplica, notas, actualizado_en, actualizado_por)
                VALUES (?, ?, 'Local', 30, 0, 0, 1, 'Auto-creado por sc-mee-asignar', datetime('now'), ?)
            """, (mee_codigo, proveedor, user))

    conn.commit()

    return jsonify({
        'ok': True,
        'sc_item_id': item_id,
        'mee_codigo': mee_codigo,
        'sku_inferido': sku,
        'componente_tipo': componente_tipo,
        'aprendizaje_nuevo': aprendizaje,
        'mensaje': (f'✅ Item asignado a {mee_codigo}'
                    + (f' · aprendido mapping {sku} → {mee_codigo} ({componente_tipo})' if aprendizaje else ' · mapping ya existía')),
    })


# ════════════════════════════════════════════════════════════════════════
# REPORTE EJECUTIVO · métricas Auto-SC + aprendizaje + valor agregado
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/reporte-ejecutivo', methods=['GET'])
def reporte_ejecutivo():
    """Métricas ejecutivas del sistema Auto-SC IA en últimos 30/90/365 días.

    Sebastián (1-may-2026): "reporte ejecutivo".

    Devuelve:
      · SCs creadas por la IA vs total (% automatización)
      · Items aprendidos (mappings nuevos en sku_mee_config)
      · MEE configurados vs sin proveedor
      · Cobertura del sistema (SKUs con mapping vs total activos)
      · Top proveedores por volumen
      · Buffer IA promedio aplicado
      · Próximas ventanas (mensual + lunes + D-20)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()
    inicio_30 = (fecha_hoy - timedelta(days=30)).isoformat()
    inicio_90 = (fecha_hoy - timedelta(days=90)).isoformat()

    rep = {'fecha': fecha_hoy.isoformat(), 'ventanas': {}}

    # 1) SCs Auto-SC IA totales
    try:
        n_total_30 = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE date(fecha) >= ?
        """, (inicio_30,)).fetchone()[0]
        n_ia_30 = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE solicitante='auto-plan-ia' AND date(fecha) >= ?
        """, (inicio_30,)).fetchone()[0]
        n_ia_90 = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE solicitante='auto-plan-ia' AND date(fecha) >= ?
        """, (inicio_90,)).fetchone()[0]
        rep['scs'] = {
            'total_30d': n_total_30,
            'ia_30d': n_ia_30,
            'ia_90d': n_ia_90,
            'pct_automatizacion_30d': round((n_ia_30 / max(n_total_30,1)) * 100, 1),
        }
    except Exception:
        rep['scs'] = {}

    # 2) Aprendizaje (mappings creados)
    try:
        n_aprendidos = c.execute("""
            SELECT COUNT(*) FROM sku_mee_config
            WHERE COALESCE(notas,'') LIKE '%Aprendido%'
              OR COALESCE(notas,'') LIKE '%Auto-mapeo%'
        """).fetchone()[0]
        n_total_mappings = c.execute("SELECT COUNT(*) FROM sku_mee_config").fetchone()[0]
        n_skus_activos = c.execute("""
            SELECT COUNT(*) FROM sku_planeacion_config
            WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
        """).fetchone()[0]
        n_skus_mapeados = c.execute("""
            SELECT COUNT(DISTINCT sku_codigo) FROM sku_mee_config WHERE aplica=1
        """).fetchone()[0]
        rep['aprendizaje'] = {
            'mappings_aprendidos': n_aprendidos,
            'mappings_totales': n_total_mappings,
            'skus_activos': n_skus_activos,
            'skus_mapeados': n_skus_mapeados,
            'pct_cobertura_skus': round((n_skus_mapeados / max(n_skus_activos,1)) * 100, 1),
        }
    except Exception:
        rep['aprendizaje'] = {}

    # 3) MEE configurados
    try:
        n_mee_total = c.execute("SELECT COUNT(*) FROM mee_lead_time_config").fetchone()[0]
        n_mee_aplica = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE aplica=1").fetchone()[0]
        n_mee_prov = c.execute("""
            SELECT COUNT(*) FROM mee_lead_time_config
            WHERE aplica=1 AND COALESCE(proveedor_principal,'') != ''
        """).fetchone()[0]
        n_mee_precio = c.execute("""
            SELECT COUNT(*) FROM mee_lead_time_config
            WHERE aplica=1 AND COALESCE(precio_unit,0) > 0
        """).fetchone()[0]
        rep['mee_config'] = {
            'total_maestro': n_mee_total,
            'aplica': n_mee_aplica,
            'con_proveedor': n_mee_prov,
            'con_precio': n_mee_precio,
            'pct_completos': round((n_mee_precio / max(n_mee_aplica,1)) * 100, 1),
        }
    except Exception:
        rep['mee_config'] = {}

    # 4) Top proveedores 90d
    try:
        rows = c.execute("""
            SELECT s.observaciones, COUNT(*) as n, SUM(COALESCE(s.valor,0)) as v
            FROM solicitudes_compra s
            WHERE s.solicitante='auto-plan-ia' AND date(s.fecha) >= ?
            GROUP BY s.observaciones
            ORDER BY n DESC LIMIT 8
        """, (inicio_90,)).fetchall()
        rep['top_proveedores'] = [
            {'observacion': (r[0] or '')[:60], 'scs': r[1], 'valor_total': round(r[2] or 0, 0)}
            for r in rows
        ]
    except Exception:
        rep['top_proveedores'] = []

    # 5) Próximas ventanas
    def _proximo_lunes(d):
        dias = (7 - d.weekday()) % 7
        if dias == 0: dias = 7
        return d + timedelta(days=dias)
    def _prox_ventana_mensual(d):
        if d.day <= 5:
            return d
        if d.month == 12:
            return date(d.year+1, 1, 1)
        return date(d.year, d.month+1, 1)
    rep['ventanas'] = {
        'proximo_lunes': _proximo_lunes(fecha_hoy).isoformat(),
        'proxima_ventana_mensual': _prox_ventana_mensual(fecha_hoy).isoformat(),
        'ventana_d20_min': (fecha_hoy + timedelta(days=15)).isoformat(),
        'ventana_d20_max': (fecha_hoy + timedelta(days=25)).isoformat(),
    }

    # 6) Items POR-ASIGNAR pendientes
    try:
        n_por_asignar = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra_items i
              JOIN solicitudes_compra s ON s.numero = i.numero
            WHERE COALESCE(i.codigo_mp,'')=''
              AND i.nombre_mp LIKE '%POR-ASIGNAR%'
              AND s.estado IN ('Pendiente','En Revision')
        """).fetchone()[0]
        rep['items_por_asignar'] = n_por_asignar
    except Exception:
        rep['items_por_asignar'] = 0

    return jsonify(rep)


# ════════════════════════════════════════════════════════════════════════
# PRE-PRODUCCIÓN · acomodo del equipo (Alejandro · Sebastián 1-may-2026)
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/pre-produccion-equipo', methods=['GET'])
def pre_produccion_equipo():
    """Resumen pre-producción próximos 7 días: producciones programadas
    con su asignación de operarios + estado de MPs/MEEs + alertas.

    Sebastián (1-may-2026): "después de planificar y solicitar tenemos que
    pasar a pre-produccion lo que queria alejandro como se acomoda el equipo".

    Por cada producción en los próximos 7 días devuelve:
      · producto + fecha + lote + cantidad_kg + area
      · operario_elaboracion + operario_envasado + operario_acondicionamiento
      · estado MPs (cuántos verificados vs pendientes vs déficit)
      · estado MEEs (envases/tapas/etiquetas listos vs faltantes)
      · alertas (operario en 2 sitios · MP en déficit · MEE faltante)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    dias = int(request.args.get('dias', 7))
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()
    fecha_max = fecha_hoy + timedelta(days=dias)

    # Producciones programadas (área desde JOIN areas_planta vía area_id)
    rows = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
               COALESCE(pp.cantidad_kg, 0),
               COALESCE(ap.codigo, '') as area_codigo,
               COALESCE(pp.estado, 'programado'),
               pp.operario_elaboracion_id, pp.operario_envasado_id,
               pp.operario_acondicionamiento_id,
               COALESCE(oe.nombre || ' ' || COALESCE(oe.apellido,''), '') as op_elab,
               COALESCE(oen.nombre || ' ' || COALESCE(oen.apellido,''), '') as op_env,
               COALESCE(oa.nombre || ' ' || COALESCE(oa.apellido,''), '') as op_acond
        FROM produccion_programada pp
          LEFT JOIN areas_planta ap ON ap.id = pp.area_id
          LEFT JOIN operarios_planta oe ON oe.id = pp.operario_elaboracion_id
          LEFT JOIN operarios_planta oen ON oen.id = pp.operario_envasado_id
          LEFT JOIN operarios_planta oa ON oa.id = pp.operario_acondicionamiento_id
        WHERE date(pp.fecha_programada) >= ?
          AND date(pp.fecha_programada) <= ?
          AND COALESCE(pp.estado, 'programado') NOT IN ('completado', 'cancelado')
        ORDER BY pp.fecha_programada ASC, pp.id ASC
    """, (fecha_hoy.isoformat(), fecha_max.isoformat())).fetchall()

    cols = ['id','producto','fecha','lotes','cantidad_kg','area','estado',
            'op_elab_id','op_env_id','op_acond_id',
            'op_elaboracion','op_envasado','op_acondicionamiento']
    producciones = []
    for r in rows:
        p = dict(zip(cols, r))
        p['op_elaboracion'] = (p['op_elaboracion'] or '').strip()
        p['op_envasado'] = (p['op_envasado'] or '').strip()
        p['op_acondicionamiento'] = (p['op_acondicionamiento'] or '').strip()
        # Estado MPs/MEEs desde checklist
        try:
            checklist = c.execute("""
                SELECT item_tipo, estado, COUNT(*) as n
                FROM produccion_checklist
                WHERE produccion_id = ?
                GROUP BY item_tipo, estado
            """, (p['id'],)).fetchall()
            mp_ok, mp_pend, mee_ok, mee_pend = 0, 0, 0, 0
            for tipo, est, n in checklist:
                tipo_low = (tipo or '').lower()
                est_low = (est or '').lower()
                es_listo = est_low in ('verificado_ok', 'recibido', 'listo', 'consumido')
                if 'envase' in tipo_low or 'tapa' in tipo_low or 'etiqueta' in tipo_low or 'caja' in tipo_low or 'serigraf' in tipo_low or 'tampograf' in tipo_low:
                    if es_listo: mee_ok += n
                    else: mee_pend += n
                else:
                    if es_listo: mp_ok += n
                    else: mp_pend += n
            p['mp_ok'] = mp_ok
            p['mp_pendientes'] = mp_pend
            p['mee_ok'] = mee_ok
            p['mee_pendientes'] = mee_pend
        except Exception:
            p['mp_ok'] = 0
            p['mp_pendientes'] = 0
            p['mee_ok'] = 0
            p['mee_pendientes'] = 0
        # Días hasta producción
        try:
            f_prod = datetime.strptime(p['fecha'][:10], '%Y-%m-%d').date()
            p['dias_hasta'] = (f_prod - fecha_hoy).days
        except Exception:
            p['dias_hasta'] = 0
        # Estado de listo
        p['listo_para_producir'] = (p['mp_pendientes'] == 0 and p['mee_pendientes'] == 0)
        producciones.append(p)

    # Detectar conflictos: mismo operario en >1 producción mismo día
    operarios_dia = {}  # (op_id, fecha) → [producciones]
    conflictos = []
    for p in producciones:
        for op_field, op_id in (('op_elaboracion', p['op_elab_id']),
                                  ('op_envasado', p['op_env_id']),
                                  ('op_acondicionamiento', p['op_acond_id'])):
            if not op_id:
                continue
            key = (op_id, p['fecha'][:10])
            if key not in operarios_dia:
                operarios_dia[key] = []
            operarios_dia[key].append({'prod_id': p['id'], 'producto': p['producto'],
                                        'rol': op_field, 'op_nombre': p[op_field]})
    for key, lst in operarios_dia.items():
        if len(lst) > 1:
            conflictos.append({
                'operario_id': key[0],
                'operario_nombre': lst[0]['op_nombre'],
                'fecha': key[1],
                'producciones': lst,
            })

    # Por operario: total carga semanal
    carga_operario = {}
    for p in producciones:
        for op_field, op_id, nombre in (
            ('elaboracion', p['op_elab_id'], p['op_elaboracion']),
            ('envasado', p['op_env_id'], p['op_envasado']),
            ('acondicionamiento', p['op_acond_id'], p['op_acondicionamiento']),
        ):
            if not op_id or not nombre:
                continue
            if op_id not in carga_operario:
                carga_operario[op_id] = {'nombre': nombre, 'producciones': 0,
                                           'kg_total': 0, 'roles': set()}
            carga_operario[op_id]['producciones'] += 1
            carga_operario[op_id]['kg_total'] += p['cantidad_kg']
            carga_operario[op_id]['roles'].add(op_field)
    carga_operario_list = []
    for op_id, info in carga_operario.items():
        carga_operario_list.append({
            'operario_id': op_id,
            'nombre': info['nombre'],
            'producciones': info['producciones'],
            'kg_total': round(info['kg_total'], 0),
            'roles': sorted(info['roles']),
        })
    carga_operario_list.sort(key=lambda x: x['producciones'], reverse=True)

    # Producciones por estado
    n_listas = sum(1 for p in producciones if p['listo_para_producir'])
    n_con_pendientes = sum(1 for p in producciones if not p['listo_para_producir'])
    n_sin_operarios = sum(1 for p in producciones
                            if not (p['op_elaboracion'] or p['op_envasado']))

    return jsonify({
        'fecha_hoy': fecha_hoy.isoformat(),
        'horizonte_dias': dias,
        'producciones': producciones,
        'kpis': {
            'total': len(producciones),
            'listas': n_listas,
            'con_pendientes': n_con_pendientes,
            'sin_operarios_asignados': n_sin_operarios,
            'conflictos_operario': len(conflictos),
        },
        'conflictos': conflictos,
        'carga_operarios': carga_operario_list,
    })


# ════════════════════════════════════════════════════════════════════════
# SYNC SHOPIFY · cron diario para mantener velocidades actualizadas
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/sync-shopify-cron', methods=['POST'])
def sync_shopify_cron():
    """Cron diario que sincroniza órdenes Shopify (Animus). Sebastián
    1-may-2026: "si sincroniza shopy".

    Acepta sesión 'compras_user' o ?clave=AUTO_PLAN_CRON_KEY (cron Render).
    Llama internamente a /api/animus/sync/shopify y registra el resultado.
    """
    clave_cron = request.args.get('clave', '') or (request.json or {}).get('clave', '')
    clave_esperada = os.environ.get('AUTO_PLAN_CRON_KEY', '')
    es_cron = bool(clave_esperada and clave_cron == clave_esperada)
    if not es_cron and 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    user = 'cron-shopify' if es_cron else session.get('compras_user', 'manual')
    conn = get_db(); c = conn.cursor()

    # Ejecutar sync Shopify usando el endpoint existente en animus.py
    # Reutilizamos la lógica directamente.
    try:
        from blueprints.animus import _cfg
        token = _cfg(conn, 'shopify_token')
        shop = _cfg(conn, 'shopify_shop')
        if not token or not shop:
            return jsonify({
                'ok': False,
                'error': 'Shopify no configurado (shopify_token/shopify_shop en config)',
                'mensaje': 'Configura las credenciales en /admin antes del cron',
            }), 400

        import urllib.request as ur
        url = f"https://{shop}/admin/api/2024-01/orders.json?status=any&limit=250"
        req = ur.Request(url, headers={"X-Shopify-Access-Token": token})
        synced = 0
        try:
            with ur.urlopen(req, timeout=30) as r:
                orders = json.loads(r.read())["orders"]
            for o in orders:
                items_sku = json.dumps([
                    {"sku": li.get("sku",""), "qty": li.get("quantity",0)}
                    for li in o.get("line_items", [])
                ])
                total_uds = sum(li.get("quantity",0) for li in o.get("line_items",[]))
                addr = o.get("billing_address") or {}
                conn.execute("""
                    INSERT OR REPLACE INTO animus_shopify_orders
                      (shopify_id, nombre, email, total, moneda, estado, estado_pago,
                       sku_items, unidades_total, ciudad, pais, creado_en, synced_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
                """, (str(o["id"]), o.get("name",""), o.get("email",""),
                      float(o.get("total_price",0)), o.get("currency","COP"),
                      o.get("fulfillment_status",""), o.get("financial_status",""),
                      items_sku, total_uds,
                      addr.get("city",""), addr.get("country_code","CO"),
                      o.get("created_at","")[:10]))
                synced += 1
            conn.commit()
        except Exception as e:
            return jsonify({'ok': False, 'error': f'Shopify API: {e}'}), 502

        # Log
        try:
            c.execute("""
                INSERT INTO auto_plan_runs
                  (ejecutado_at, ejecutado_por, tipo, horizonte_dias,
                   producciones_creadas, compras_creadas, alertas_criticas,
                   emails_enviados, error, payload_json)
                VALUES (?, ?, 'sync_shopify_cron', 0, 0, 0, 0, 0, NULL, ?)
            """, (datetime.now().isoformat(), user,
                  json.dumps({'orders_synced': synced})))
            conn.commit()
        except Exception:
            pass

        return jsonify({
            'ok': True,
            'orders_synced': synced,
            'mensaje': f'✅ {synced} órdenes Shopify sincronizadas',
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ════════════════════════════════════════════════════════════════════════
# MULTI-CRON STATUS · monitoreo de jobs internos
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/cron-jobs-status', methods=['GET'])
def cron_jobs_status():
    """Status de los 5 jobs internos del multi-cron.
    Sebastián 1-may-2026: cron interno sin Render externos.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    try:
        from blueprints.auto_plan_jobs import JOBS_SCHEDULE
    except Exception:
        return jsonify({'error': 'multi-cron no inicializado'}), 500

    jobs_info = []
    for job_name, hora, minuto, dias_sem, dias_mes, callable_name in JOBS_SCHEDULE:
        try:
            ult = c.execute("""
                SELECT ejecutado_at, ok, duracion_ms, resultado_json, error
                FROM cron_jobs_runs
                WHERE job_name = ?
                ORDER BY id DESC LIMIT 1
            """, (job_name,)).fetchone()
        except Exception:
            ult = None
        # Schedule legible
        if dias_mes:
            sched = f"día {','.join(map(str, dias_mes))} mes a las {hora:02d}:{minuto:02d}"
        elif dias_sem:
            nombres = ['L','M','Mi','J','V','S','D']
            sched = f"{','.join(nombres[d] for d in dias_sem)} a las {hora:02d}:{minuto:02d}"
        else:
            sched = f"diario {hora:02d}:{minuto:02d}"
        jobs_info.append({
            'job_name': job_name,
            'schedule': sched,
            'ultima_ejecucion_at': ult[0] if ult else None,
            'ultima_ok': bool(ult[1]) if ult else None,
            'ultima_duracion_ms': ult[2] if ult else None,
            'ultima_resultado': ult[3] if ult else None,
            'ultima_error': ult[4] if ult else None,
        })
    return jsonify({
        'jobs': jobs_info,
        'total': len(jobs_info),
        'mensaje': 'Multi-cron interno corriendo · revisa cada 5 min',
    })


# ════════════════════════════════════════════════════════════════════════
# MI DÍA · vista por operario (Sebastián 1-may-2026)
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/mi-dia', methods=['GET'])
def mi_dia():
    """Vista 'Mi día' por operario: tareas próximos 7 días.
    Sebastián 1-may-2026: "preproducción que es que todos sepan que les
    toca hacer". Reglas Alejandro (memoria):
      · L/Mi/V → producir
      · Ma/J → acondicionar/conteo/limpieza
      · Limpieza profunda obligatoria si producto previo tiene pigmento alto
        (riesgo_arrastre_pct >= 50)

    Query: ?operario_id=X o ?usuario=catalina (auto-detect del operario)
           ?dias=7 (default 7)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    operario_id = request.args.get('operario_id', type=int)
    usuario = (request.args.get('usuario') or session.get('compras_user') or '').strip()
    dias = int(request.args.get('dias', 7))

    # Resolver operario_id si solo viene usuario
    if not operario_id and usuario:
        row = c.execute("""
            SELECT id, nombre, apellido, rol_predeterminado
            FROM operarios_planta
            WHERE LOWER(nombre) LIKE LOWER(?)
               OR LOWER(nombre || ' ' || COALESCE(apellido,'')) LIKE LOWER(?)
            LIMIT 1
        """, (f'%{usuario}%', f'%{usuario}%')).fetchone()
        if row:
            operario_id = row[0]

    if not operario_id:
        # Listar operarios para que el frontend deje al usuario elegir
        rows = c.execute("""
            SELECT id, nombre || ' ' || COALESCE(apellido,'') as nombre,
                   rol_predeterminado, activo
            FROM operarios_planta WHERE COALESCE(activo,1)=1
            ORDER BY nombre
        """).fetchall()
        return jsonify({
            'sin_operario_resuelto': True,
            'usuario_query': usuario,
            'operarios_disponibles': [
                {'id': r[0], 'codigo': '', 'nombre': (r[1] or '').strip(),
                 'rol': r[2] or '', 'activo': bool(r[3])}
                for r in rows
            ],
        })

    # Datos del operario
    op = c.execute("""
        SELECT id, nombre, apellido, rol_predeterminado,
               COALESCE(fija_en_dispensacion,0), COALESCE(es_jefe_produccion,0)
        FROM operarios_planta WHERE id = ?
    """, (operario_id,)).fetchone()
    if not op:
        return jsonify({'error': 'Operario no encontrado'}), 404
    op_info = {
        'id': op[0], 'codigo': '',
        'nombre': (op[1] or '').strip() + ' ' + (op[2] or '').strip(),
        'rol_predeterminado': op[3] or '',
        'fija_en_dispensacion': bool(op[4]),
        'es_jefe_produccion': bool(op[5]),
    }

    fecha_hoy = datetime.now().date()
    fecha_max = fecha_hoy + timedelta(days=dias)
    nombres_dia = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    DIAS_PRODUCCION = (0, 2, 4)  # L, Mi, V
    DIAS_ACONDICIONAR = (1, 3)   # Ma, J

    # Tareas por día
    dias_data = []
    for d_offset in range(dias):
        f = fecha_hoy + timedelta(days=d_offset)
        wd = f.weekday()
        dia_label = 'PRODUCCIÓN' if wd in DIAS_PRODUCCION else (
                    'ACONDICIONAR/CONTEO' if wd in DIAS_ACONDICIONAR else 'FIN DE SEMANA')
        es_lab = wd < 5

        # Producciones donde participa este operario
        prods = c.execute("""
            SELECT pp.id, pp.producto, pp.lotes, COALESCE(pp.cantidad_kg, 0),
                   ap.codigo as area, pp.estado,
                   CASE
                     WHEN pp.operario_dispensacion_id = ? THEN 'dispensacion'
                     WHEN pp.operario_elaboracion_id = ? THEN 'elaboracion'
                     WHEN pp.operario_envasado_id = ? THEN 'envasado'
                     WHEN pp.operario_acondicionamiento_id = ? THEN 'acondicionamiento'
                     ELSE 'apoyo'
                   END as rol_en_prod,
                   pr.tiene_pigmento, pr.color_descripcion, pr.riesgo_arrastre_pct
            FROM produccion_programada pp
              LEFT JOIN areas_planta ap ON ap.id = pp.area_id
              LEFT JOIN producto_perfil_riesgo pr
                ON UPPER(TRIM(pr.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE date(pp.fecha_programada) = ?
              AND COALESCE(pp.estado, 'programado') NOT IN ('completado', 'cancelado')
              AND (pp.operario_dispensacion_id = ? OR pp.operario_elaboracion_id = ?
                   OR pp.operario_envasado_id = ? OR pp.operario_acondicionamiento_id = ?)
        """, (operario_id, operario_id, operario_id, operario_id, f.isoformat(),
              operario_id, operario_id, operario_id, operario_id)).fetchall()

        prods_data = [{
            'id': p[0], 'producto': p[1], 'lotes': p[2], 'kg': p[3],
            'area': p[4] or '', 'estado': p[5], 'rol': p[6],
            'pigmento': bool(p[7]) if p[7] is not None else False,
            'color_descripcion': p[8] or '',
            'riesgo_arrastre_pct': p[9] or 0,
        } for p in prods]

        # Limpiezas profundas asignadas a este operario
        try:
            limps = c.execute("""
                SELECT id, area_codigo, asignado_a, estado, razon_asignacion
                FROM limpieza_profunda_calendario
                WHERE date(fecha) = ?
                  AND (asignado_a = ? OR LOWER(asignado_a) = LOWER(?))
            """, (f.isoformat(), op_info['nombre'], op_info['nombre'])).fetchall()
            limps_data = [{'id': l[0], 'area': l[1], 'asignado_a': l[2],
                            'estado': l[3], 'razon': l[4] or ''} for l in limps]
        except Exception:
            limps_data = []

        # Conteos cíclicos asignados
        try:
            conts = c.execute("""
                SELECT id, material_id, material_nombre, estado, categoria_abc
                FROM conteo_ciclico_calendario
                WHERE date(fecha) = ?
                  AND (asignado_a = ? OR LOWER(asignado_a) = LOWER(?))
            """, (f.isoformat(), op_info['nombre'], op_info['nombre'])).fetchall()
            cont_data = [{'id': c_[0], 'material_id': c_[1], 'material': c_[2],
                           'estado': c_[3], 'abc': c_[4] or ''} for c_ in conts]
        except Exception:
            cont_data = []

        # Limpieza obligatoria si producción anterior tiene pigmento alto
        alerta_limpieza = None
        if prods_data:
            for p in prods_data:
                if p['riesgo_arrastre_pct'] >= 50:
                    alerta_limpieza = (f"⚠️ Limpieza profunda OBLIGATORIA antes de "
                                        f"siguiente producción (anterior tuvo pigmento "
                                        f"{p['color_descripcion']}, riesgo {p['riesgo_arrastre_pct']}%)")
                    break

        dias_data.append({
            'fecha': f.isoformat(),
            'nombre_dia': nombres_dia[wd],
            'tipo_dia': dia_label,
            'es_laboral': es_lab,
            'producciones': prods_data,
            'limpiezas': limps_data,
            'conteos': cont_data,
            'alerta_limpieza': alerta_limpieza,
            'total_tareas': len(prods_data) + len(limps_data) + len(cont_data),
        })

    # Resumen
    total_tareas = sum(d['total_tareas'] for d in dias_data)
    total_kg = sum(p['kg'] for d in dias_data for p in d['producciones'])

    return jsonify({
        'operario': op_info,
        'fecha_hoy': fecha_hoy.isoformat(),
        'horizonte_dias': dias,
        'dias': dias_data,
        'resumen': {
            'total_tareas': total_tareas,
            'kg_total_semana': round(total_kg, 0),
            'dias_con_actividad': sum(1 for d in dias_data if d['total_tareas'] > 0),
        },
    })


@bp.route('/api/planta/semana-produccion', methods=['GET'])
def semana_produccion():
    """Vista CLARA para el jefe de producción · IA asignó todo,
    el operario solo da click para avanzar.
    Sebastián 1-may-2026: 'que sea aún más sencillo · qué produce esta semana
    le salga lunes tal · dónde queda cada operario · solicitar limpieza marque
    limpiado/sucio · todo en tiempo real · pero que todo lo asigne la IA'.

    Devuelve L-V (lunes próximo si hoy es fin de semana) con:
      - Por cada día: producciones programadas, área asignada por IA,
        operarios rotados por IA, estado live, limpiezas
      - Si una producción NO tiene área/operario → flag (IA debe asignar)
      - Acciones disponibles según estado: iniciar / terminar / marcar limpia
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()

    # Encontrar lunes de esta semana (si hoy es fin de semana, próximo lunes)
    base = fecha_hoy
    if base.weekday() >= 5:  # sáb/dom
        base = base + timedelta(days=(7 - base.weekday()))
    else:
        base = base - timedelta(days=base.weekday())

    dias_semana = [base + timedelta(days=i) for i in range(5)]
    nombres_dia = ['LUN', 'MAR', 'MIÉ', 'JUE', 'VIE']

    # Pre-cargar eventos del Calendar para los próximos 14d (para mergear cuando
    # produccion_programada esté vacía o tenga menos que el Calendar)
    calendar_eventos_por_fecha = {}
    try:
        cal_events = _calendar_events_cached() or []
        skus_aliases = {}
        try:
            for sku_n, alias_csv in c.execute("""
                SELECT producto_nombre, COALESCE(alias_calendar,'')
                FROM sku_planeacion_config
                WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
            """).fetchall():
                skus_aliases[sku_n] = alias_csv
        except Exception:
            pass
        for ev in cal_events:
            try:
                f_ev = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            except Exception:
                continue
            if f_ev < dias_semana[0] or f_ev > dias_semana[-1]:
                continue
            # Match producto
            producto_match = None
            best_score = 0
            for prod_n, alias_csv in skus_aliases.items():
                try:
                    score = _match_producto_evento(prod_n, alias_csv,
                                                     ev.get('titulo'), ev.get('descripcion',''))
                    if score >= 60 and score > best_score:
                        best_score = score
                        producto_match = prod_n
                except Exception:
                    continue
            kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion','')) or 0
            calendar_eventos_por_fecha.setdefault(f_ev.isoformat(), []).append({
                'titulo': ev.get('titulo', ''),
                'producto_match': producto_match,
                'kg': kg,
                'desde_calendar': True,
            })
    except Exception:
        pass

    dias_data = []
    for i, fecha in enumerate(dias_semana):
        es_hoy = (fecha == fecha_hoy)
        wd = fecha.weekday()
        es_lmv = wd in (0, 2, 4)  # lunes, miércoles, viernes
        tipo_dia = 'PRODUCCIÓN' if es_lmv else 'ACOND/CONTEO'

        # Producciones programadas ese día (desde produccion_programada)
        prods_rows = c.execute("""
            SELECT pp.id, pp.producto, pp.lotes, COALESCE(pp.cantidad_kg, 0),
                   pp.estado, pp.area_id,
                   ap.codigo as area_codigo, ap.nombre as area_nombre, ap.estado as area_estado,
                   ap_env.codigo as env_codigo,
                   o1.id, o1.nombre || ' ' || COALESCE(o1.apellido,'') as op_disp,
                   o2.id, o2.nombre || ' ' || COALESCE(o2.apellido,'') as op_elab,
                   o3.id, o3.nombre || ' ' || COALESCE(o3.apellido,'') as op_env,
                   o4.id, o4.nombre || ' ' || COALESCE(o4.apellido,'') as op_acon,
                   pp.inicio_real_at, pp.fin_real_at
            FROM produccion_programada pp
              LEFT JOIN areas_planta ap ON ap.id = pp.area_id
              LEFT JOIN areas_planta ap_env ON ap_env.id = pp.area_envasado_id
              LEFT JOIN operarios_planta o1 ON o1.id = pp.operario_dispensacion_id
              LEFT JOIN operarios_planta o2 ON o2.id = pp.operario_elaboracion_id
              LEFT JOIN operarios_planta o3 ON o3.id = pp.operario_envasado_id
              LEFT JOIN operarios_planta o4 ON o4.id = pp.operario_acondicionamiento_id
            WHERE date(pp.fecha_programada) = ?
              AND COALESCE(pp.estado, 'programado') != 'cancelado'
            ORDER BY pp.id
        """, (fecha.isoformat(),)).fetchall()

        prods = []
        for r in prods_rows:
            (pid, producto, lotes, kg, estado, area_id, area_cod, area_nom, area_est,
             env_cod, _, op_disp, _, op_elab, _, op_env, _, op_acon,
             inicio_at, fin_at) = r
            # Bloqueado lunes 7am? (lookup separado · col puede no existir aún)
            bloqueado = False
            try:
                bl_row = c.execute("SELECT bloqueado_at FROM produccion_programada WHERE id=?", (pid,)).fetchone()
                bloqueado = bool(bl_row and bl_row[0])
            except Exception:
                pass
            estado = (estado or 'programado').strip()
            ya_asignada = bool(area_id and (op_disp or op_elab or op_env))
            # Acción siguiente según estado
            if estado == 'completado':
                accion = None; accion_label = '✅ Completada'
            elif estado in ('en_proceso', 'iniciado'):
                accion = 'terminar'; accion_label = '✓ Terminar'
            elif ya_asignada:
                accion = 'iniciar'; accion_label = '▶ Iniciar'
            else:
                accion = 'asignar_ia'; accion_label = '🤖 IA asignar'
            prods.append({
                'id': pid, 'producto': producto, 'lotes': lotes, 'kg': kg,
                'estado': estado,
                'bloqueado': bloqueado,
                'area': {'codigo': area_cod, 'nombre': area_nom, 'estado': area_est},
                'envasado': env_cod,
                'operarios': {
                    'dispensacion': (op_disp or '').strip(),
                    'elaboracion': (op_elab or '').strip(),
                    'envasado': (op_env or '').strip(),
                    'acondicionamiento': (op_acon or '').strip(),
                },
                'ya_asignada': ya_asignada,
                'accion': accion, 'accion_label': accion_label,
                'inicio_real_at': inicio_at, 'fin_real_at': fin_at,
            })

        # MERGE: agregar eventos del Calendar que NO tienen contraparte en DB
        # (para que aparezcan aunque la sync calendar→produccion_programada no
        # haya corrido aún)
        productos_en_db = {p['producto'] for p in prods if p.get('producto')}
        for cev in calendar_eventos_por_fecha.get(fecha.isoformat(), []):
            prod_n = cev.get('producto_match') or cev.get('titulo','').split('–')[0].strip()
            if not prod_n: continue
            # Si ya está en DB, skip (evita duplicados)
            if any(prod_n.upper() == (p.get('producto') or '').upper() for p in prods):
                continue
            prods.append({
                'id': None, 'producto': prod_n,
                'lotes': 1, 'kg': cev.get('kg') or 0,
                'estado': 'calendar_sin_sync',
                'area': {'codigo': '', 'nombre': '', 'estado': ''},
                'envasado': '',
                'operarios': {'dispensacion':'','elaboracion':'','envasado':'','acondicionamiento':''},
                'ya_asignada': False,
                'accion': 'sincronizar', 'accion_label': '🔄 Sync Calendar',
                'desde_calendar': True,
                'titulo_calendar': cev.get('titulo', '')[:60],
            })

        # Limpiezas programadas ese día
        limpiezas = []
        try:
            limp_rows = c.execute("""
                SELECT id, area_codigo, asignado_a, estado, razon_asignacion
                FROM limpieza_profunda_calendario
                WHERE date(fecha) = ?
                  AND estado IN ('pendiente','asignada','en_proceso')
                ORDER BY id
            """, (fecha.isoformat(),)).fetchall()
            for lr in limp_rows:
                limpiezas.append({
                    'id': lr[0], 'area': lr[1], 'asignado_a': lr[2] or '',
                    'estado': lr[3], 'razon': (lr[4] or '')[:60],
                    'accion': 'marcar_limpia' if lr[3] != 'completada' else None,
                })
        except Exception:
            pass

        dias_data.append({
            'fecha': fecha.isoformat(),
            'nombre_dia': nombres_dia[i],
            'es_hoy': es_hoy,
            'tipo_dia': tipo_dia,
            'es_laboral': True,
            'producciones': prods,
            'limpiezas': limpiezas,
            'total_kg': sum(p['kg'] for p in prods),
            'total_lotes': sum(p['lotes'] or 0 for p in prods),
            'sin_asignar': sum(1 for p in prods if not p['ya_asignada']),
            'sin_sync': sum(1 for p in prods if p.get('desde_calendar')),
        })

    # Estado salas en vivo (mini)
    salas = []
    try:
        rows = c.execute("""
            SELECT codigo, nombre, estado FROM areas_planta
            WHERE activo=1 AND tipo='produccion'
            ORDER BY orden, codigo
        """).fetchall()
        salas = [{'codigo': r[0], 'nombre': r[1], 'estado': r[2] or 'libre'} for r in rows]
    except Exception:
        pass

    # KPIs semana
    total_prods = sum(len(d['producciones']) for d in dias_data)
    total_kg = sum(d['total_kg'] for d in dias_data)
    sin_asignar_total = sum(d['sin_asignar'] for d in dias_data)
    total_limpiezas = sum(len(d['limpiezas']) for d in dias_data)

    return jsonify({
        'fecha_hoy': fecha_hoy.isoformat(),
        'semana_inicio': dias_semana[0].isoformat(),
        'semana_fin': dias_semana[-1].isoformat(),
        'dias': dias_data,
        'salas': salas,
        'kpis': {
            'total_producciones_semana': total_prods,
            'total_kg_semana': round(total_kg, 0),
            'sin_asignar_ia': sin_asignar_total,
            'total_limpiezas': total_limpiezas,
        },
    })


@bp.route('/api/planta/accion-rapida', methods=['POST'])
def planta_accion_rapida():
    """Acciones rápidas del operario en producciones/limpiezas.
    Sebastián 1-may-2026: "marque limpiado/sucio · iniciar/terminar".

    Body: {tipo, id, ...}
    Tipos:
      - iniciar_produccion: {tipo:'iniciar_produccion', produccion_id}
      - terminar_produccion: {tipo:'terminar_produccion', produccion_id, unidades_envasadas?}
      - marcar_limpia: {tipo:'marcar_limpia', limpieza_id} → área pasa a 'libre'
      - marcar_sucia: {tipo:'marcar_sucia', area_codigo}
      - asignar_ia: {tipo:'asignar_ia', produccion_id} → fuerza auto-asignación IA
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    tipo = (d.get('tipo') or '').strip()
    conn = get_db(); c = conn.cursor()

    if tipo == 'iniciar_produccion':
        pid = d.get('produccion_id')
        if not pid: return jsonify({'error': 'produccion_id requerido'}), 400
        try:
            row = c.execute("SELECT area_id, estado FROM produccion_programada WHERE id=?", (pid,)).fetchone()
            if not row: return jsonify({'error': 'producción no existe'}), 404
            if row[1] in ('en_proceso', 'iniciado'):
                return jsonify({'ok': True, 'mensaje': 'Ya estaba iniciada'})
            if row[1] == 'completado':
                return jsonify({'error': 'Ya completada'}), 400
            c.execute("""
                UPDATE produccion_programada
                  SET estado='en_proceso', inicio_real_at=datetime('now')
                WHERE id=?
            """, (pid,))
            # Marcar área como ocupada
            if row[0]:
                c.execute("UPDATE areas_planta SET estado='ocupada' WHERE id=? AND estado IN ('libre','limpiando')", (row[0],))
            conn.commit()
            return jsonify({'ok': True, 'mensaje': '▶ Producción iniciada'})
        except Exception as e:
            try: conn.rollback()
            except: pass
            return jsonify({'error': str(e)}), 500

    elif tipo == 'terminar_produccion':
        pid = d.get('produccion_id')
        if not pid: return jsonify({'error': 'produccion_id requerido'}), 400
        # Llamar al endpoint existente de completar (que ya hace descuento + limpieza auto)
        try:
            from blueprints.programacion import prog_completar_evento
            # No podemos llamar directo porque hace request.get_json() interno
            # Mejor hacer el flujo aquí simplificado:
            row = c.execute("""
                SELECT id, producto, fecha_programada, area_id
                FROM produccion_programada WHERE id=?
            """, (pid,)).fetchone()
            if not row: return jsonify({'error': 'producción no existe'}), 404
            c.execute("""
                UPDATE produccion_programada
                  SET estado='completado', fin_real_at=datetime('now')
                WHERE id=?
            """, (pid,))
            # Marcar área sucia + crear limpieza auto
            if row[3]:
                c.execute("UPDATE areas_planta SET estado='sucia' WHERE id=?", (row[3],))
                try:
                    from blueprints.programacion import _crear_limpieza_post_produccion
                    fecha_iso = row[2][:10] if row[2] else datetime.now().date().isoformat()
                    area_row = c.execute("SELECT codigo FROM areas_planta WHERE id=?", (row[3],)).fetchone()
                    if area_row:
                        _crear_limpieza_post_produccion(c, row[3], area_row[0], fecha_iso,
                                                          row[1], '', user)
                except Exception:
                    pass
            conn.commit()
            return jsonify({'ok': True, 'mensaje': '✓ Producción terminada · área marcada sucia · limpieza programada'})
        except Exception as e:
            try: conn.rollback()
            except: pass
            return jsonify({'error': str(e)}), 500

    elif tipo == 'marcar_limpia':
        lid = d.get('limpieza_id')
        if not lid: return jsonify({'error': 'limpieza_id requerido'}), 400
        try:
            row = c.execute("SELECT area_codigo FROM limpieza_profunda_calendario WHERE id=?", (lid,)).fetchone()
            if not row: return jsonify({'error': 'limpieza no existe'}), 404
            c.execute("""
                UPDATE limpieza_profunda_calendario
                  SET estado='completada', terminado_at=datetime('now'), terminado_por=?
                WHERE id=?
            """, (user, lid))
            # Área pasa a libre
            c.execute("UPDATE areas_planta SET estado='libre', ultima_limpieza_profunda=datetime('now') WHERE codigo=?", (row[0],))
            conn.commit()
            return jsonify({'ok': True, 'mensaje': f'✓ Área {row[0]} limpia · disponible para producción'})
        except Exception as e:
            try: conn.rollback()
            except: pass
            return jsonify({'error': str(e)}), 500

    elif tipo == 'marcar_sucia':
        codigo = (d.get('area_codigo') or '').strip()
        if not codigo: return jsonify({'error': 'area_codigo requerido'}), 400
        c.execute("UPDATE areas_planta SET estado='sucia' WHERE codigo=?", (codigo,))
        conn.commit()
        return jsonify({'ok': True, 'mensaje': f'Área {codigo} marcada sucia'})

    elif tipo == 'asignar_ia':
        pid = d.get('produccion_id')
        if not pid: return jsonify({'error': 'produccion_id requerido'}), 400
        try:
            from blueprints.programacion import _auto_asignar_produccion
            res = _auto_asignar_produccion(c, pid, f'manual-{user}')
            if res.get('ok'):
                conn.commit()
                return jsonify({'ok': True, 'mensaje': '🤖 IA asignó: ' + ' · '.join(res.get('cambios', []))})
            else:
                return jsonify({'ok': False, 'error': res.get('error')}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    elif tipo == 'sincronizar':
        # Sebastián 1-may-2026: sync del Calendar a produccion_programada
        # + auto-asignar IA inmediatamente. Procesa TODOS los pendientes
        # (no solo uno) porque vienen del Calendar feed.
        try:
            from blueprints.programacion import (
                _calendar_events_cached, _auto_asignar_produccion
            )
            cal_events = _calendar_events_cached(force_refresh=True) or []
            skus_aliases = {}
            for sku_n, alias_csv in c.execute("""
                SELECT producto_nombre, COALESCE(alias_calendar, '')
                FROM sku_planeacion_config
                WHERE activo = 1
                  AND COALESCE(estado, 'activo') NOT IN ('descontinuado', 'pausado')
            """).fetchall():
                skus_aliases[sku_n] = alias_csv
            insertados = 0
            asignados = 0
            for ev in cal_events:
                try:
                    f_ev = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
                except Exception:
                    continue
                if f_ev < datetime.now().date():
                    continue
                # Match producto
                producto_match = None
                best_score = 0
                for prod_n, alias_csv in skus_aliases.items():
                    try:
                        score = _match_producto_evento(prod_n, alias_csv,
                                                         ev.get('titulo'), ev.get('descripcion', ''))
                        if score >= 60 and score > best_score:
                            best_score = score
                            producto_match = prod_n
                    except Exception:
                        continue
                if not producto_match:
                    continue
                kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', '')) or 0
                # Skip si ya existe
                exists = c.execute("""
                    SELECT id FROM produccion_programada
                    WHERE producto = ? AND date(fecha_programada) = ?
                """, (producto_match, f_ev.isoformat())).fetchone()
                if exists:
                    pid_existente = exists[0]
                    # Si no tiene área/operarios, auto-asignar
                    row = c.execute("SELECT area_id, operario_dispensacion_id FROM produccion_programada WHERE id=?", (pid_existente,)).fetchone()
                    if row and not row[0] and not row[1]:
                        res = _auto_asignar_produccion(c, pid_existente, f'sync-{user}')
                        if res.get('ok'):
                            asignados += 1
                    continue
                # Insertar y asignar
                cur_ins = c.execute("""
                    INSERT INTO produccion_programada
                      (producto, fecha_programada, lotes, cantidad_kg,
                       estado, observaciones, origen)
                    VALUES (?, ?, 1, ?, 'programado', ?, 'calendar')
                """, (producto_match, f_ev.isoformat(), kg,
                      f'[sync manual] {(ev.get("titulo") or "")[:200]}'))
                new_id = cur_ins.lastrowid
                insertados += 1
                if kg > 0 and new_id:
                    res = _auto_asignar_produccion(c, new_id, f'sync-{user}')
                    if res.get('ok'):
                        asignados += 1
            conn.commit()
            return jsonify({
                'ok': True,
                'mensaje': f'🔄 Sync Calendar: {insertados} producciones nuevas · {asignados} asignadas por IA',
            })
        except Exception as e:
            try: conn.rollback()
            except: pass
            return jsonify({'error': str(e)}), 500

    return jsonify({'error': f'tipo desconocido: {tipo}'}), 400


@bp.route('/api/planta/estado-solicitudes', methods=['GET'])
def estado_solicitudes():
    """Estado del Plan: qué se solicitó vs qué falta solicitar.
    Sebastián 1-may-2026: 'Plan: yo ya le di solicitar todo · debería decir
    qué ya se solicitó y que no'.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    # SCs creadas por IA en último mes (MP + MEE)
    sc_mp = 0; sc_mee = 0; sc_etq = 0; sc_d20 = 0
    try:
        sc_mp = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE solicitante='auto-plan-ia'
              AND categoria='Materia Prima'
              AND date(fecha) >= date('now','-30 days')
        """).fetchone()[0]
        sc_mee = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE solicitante='auto-plan-ia'
              AND categoria='Material de Empaque'
              AND date(fecha) >= date('now','-30 days')
        """).fetchone()[0]
        sc_etq = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE date(fecha) >= date('now','-14 days')
              AND observaciones LIKE '%etiqueta%'
        """).fetchone()[0]
        sc_d20 = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE date(fecha) >= date('now','-30 days')
              AND categoria='Servicios'
              AND observaciones LIKE '%D-20%'
        """).fetchone()[0]
    except Exception:
        pass

    # Estado del último workflow lunes
    ult_lunes = None
    try:
        row = c.execute("""
            SELECT ejecutado_at, fecha_lunes, producciones_bloqueadas,
                   sincronizadas, asignadas, limpiezas_creadas, email_enviado
            FROM workflow_lunes_log
            ORDER BY id DESC LIMIT 1
        """).fetchone()
        if row:
            ult_lunes = {
                'ejecutado_at': row[0], 'fecha_lunes': row[1],
                'bloqueadas': row[2], 'sincronizadas': row[3],
                'asignadas': row[4], 'limpiezas_creadas': row[5],
                'email_enviado': bool(row[6]),
            }
    except Exception:
        pass

    # Calcular si debería ejecutarse ahora (lunes 7am o pendiente)
    from datetime import datetime as _dt, timedelta as _td
    fecha_hoy = _dt.now().date()
    base = fecha_hoy
    while base.weekday() != 0:
        base -= _td(days=1)
    lunes_actual = base.isoformat()
    workflow_de_esta_semana = ult_lunes and ult_lunes.get('fecha_lunes') == lunes_actual

    # Calcular preview de SCs pendientes (qué falta solicitar)
    pendientes = {'mp_30d': 0, 'mee_30d': 0}
    try:
        from blueprints.auto_plan import _calcular_auto_sc, _calcular_auto_sc_mee
        plan_mp = _calcular_auto_sc(conn, modo='mensual')
        pendientes['mp_30d'] = plan_mp.get('kpis', {}).get('total_items', 0)
        plan_mee = _calcular_auto_sc_mee(conn, modo='mensual')
        pendientes['mee_30d'] = plan_mee.get('kpis', {}).get('total_items', 0)
    except Exception:
        pass

    return jsonify({
        'fecha_hoy': fecha_hoy.isoformat(),
        'lunes_actual': lunes_actual,
        'workflow_lunes_ejecutado_esta_semana': bool(workflow_de_esta_semana),
        'ultimo_workflow_lunes': ult_lunes,
        'solicitado_ultimo_mes': {
            'mp': sc_mp, 'mee': sc_mee, 'etiquetas': sc_etq, 'd20': sc_d20,
            'total': sc_mp + sc_mee + sc_etq + sc_d20,
        },
        'pendiente_solicitar': pendientes,
    })


@bp.route('/api/planta/ejecutar-lunes-7am', methods=['POST'])
def ejecutar_lunes_7am():
    """Ejecuta manualmente el workflow del lunes 7am (botón Sebastián).
    Útil cuando se quiere disparar antes/después o re-ejecutar."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    try:
        from blueprints.auto_plan_jobs import job_lunes_7am_workflow
        from flask import current_app
        ok, resultado, _ = job_lunes_7am_workflow(current_app)
        return jsonify({'ok': ok, 'resultado': resultado,
                          'mensaje': f'⭐ Workflow lunes 7am ejecutado · {resultado.get("bloqueadas", 0)} producciones bloqueadas'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/planta/desbloquear-produccion/<int:pid>', methods=['POST'])
def desbloquear_produccion(pid):
    """Desbloquear una producción específica (CEO/Alejandro)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    c.execute("""
        UPDATE produccion_programada SET bloqueado_at=NULL, bloqueado_por='' WHERE id=?
    """, (pid,))
    conn.commit()
    return jsonify({'ok': True, 'mensaje': f'Producción #{pid} desbloqueada'})


@bp.route('/api/planta/health-check', methods=['GET'])
def planta_health_check():
    """Health check del sistema completo · Sebastián 1-may-2026:
    'meta que todo se programe solo automático'.

    Devuelve estado de cada componente con verde/amarillo/rojo + sugerencias.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()
    items = []  # [{categoria, status, nombre, valor, sugerencia}]

    # 1) Cron auto-plan habilitado
    try:
        r = c.execute("SELECT habilitado, errores_consecutivos FROM auto_plan_cron_state WHERE id=1").fetchone()
        habilitado = bool(r[0]) if r else False
        items.append({
            'categoria': 'cron',
            'nombre': 'Auto-plan cron habilitado',
            'status': 'ok' if habilitado else 'error',
            'valor': 'sí' if habilitado else 'no',
            'sugerencia': '' if habilitado else 'Habilita en /api/auto-plan/cron/toggle',
        })
        if r and r[1] > 0:
            items.append({
                'categoria': 'cron',
                'nombre': 'Errores consecutivos cron',
                'status': 'warn' if r[1] < 3 else 'error',
                'valor': str(r[1]),
                'sugerencia': 'Revisa logs si >3',
            })
    except Exception as e:
        items.append({'categoria':'cron','nombre':'Auto-plan cron state','status':'error','valor':str(e),'sugerencia':''})

    # 2) Multi-cron · 6 jobs ejecutándose
    try:
        from blueprints.auto_plan_jobs import JOBS_SCHEDULE
        for job_name, h, m, _, _, _ in JOBS_SCHEDULE:
            ult = c.execute("""
                SELECT ejecutado_at, ok FROM cron_jobs_runs
                WHERE job_name=? ORDER BY id DESC LIMIT 1
            """, (job_name,)).fetchone()
            sched = f'{h:02d}:{m:02d}'
            if not ult:
                items.append({'categoria':'multi-cron','nombre':f'Job {job_name}','status':'warn',
                                'valor':f'nunca · {sched}','sugerencia':'Esperando primera ejecución'})
            else:
                ago = (datetime.now() - datetime.fromisoformat(ult[0])).total_seconds() / 3600
                if ult[1]:
                    items.append({'categoria':'multi-cron','nombre':f'Job {job_name}','status':'ok',
                                    'valor':f'hace {ago:.1f}h ({sched})','sugerencia':''})
                else:
                    items.append({'categoria':'multi-cron','nombre':f'Job {job_name}','status':'error',
                                    'valor':'último FALLÓ','sugerencia':'Revisar cron_jobs_runs.error'})
    except Exception:
        pass

    # 3) Calendar conectado
    try:
        from blueprints.programacion import _fetch_calendar_events
        result = _fetch_calendar_events(days_ahead=60) or {}
        eventos = result.get('events') or []
        if result.get('source') == 'none':
            items.append({'categoria':'calendar','nombre':'Calendar conectado','status':'error',
                            'valor':'no configurado','sugerencia':'Configura GCAL_ICAL_URL en Render'})
        elif result.get('error'):
            items.append({'categoria':'calendar','nombre':'Calendar fetch','status':'error',
                            'valor':result['error'][:60],'sugerencia':'Revisa URL iCal'})
        elif not eventos:
            items.append({'categoria':'calendar','nombre':'Eventos próximos 60d','status':'warn',
                            'valor':'0 eventos','sugerencia':'Crea eventos en Calendar Producciones'})
        else:
            items.append({'categoria':'calendar','nombre':'Eventos próximos 60d','status':'ok',
                            'valor':f'{len(eventos)} eventos','sugerencia':''})
    except Exception as e:
        items.append({'categoria':'calendar','nombre':'Calendar','status':'error','valor':str(e)[:60],'sugerencia':''})

    # 4) Email destinatarios
    try:
        n = c.execute("SELECT COUNT(*) FROM email_destinatarios_config WHERE activo=1 AND email != ''").fetchone()[0]
        items.append({'categoria':'email','nombre':'Destinatarios email','status':'ok' if n > 0 else 'warn',
                        'valor':f'{n} configurados','sugerencia':'Configura emails en /admin para que lleguen las alertas' if n == 0 else ''})
    except Exception:
        pass

    # 5) MEE config completa
    try:
        n_total = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE aplica=1").fetchone()[0]
        n_prov = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE aplica=1 AND COALESCE(proveedor_principal,'')!=''").fetchone()[0]
        n_precio = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE aplica=1 AND COALESCE(precio_unit,0)>0").fetchone()[0]
        pct_prov = (n_prov/max(n_total,1))*100
        pct_precio = (n_precio/max(n_total,1))*100
        items.append({'categoria':'mee','nombre':'MEE con proveedor','status':'ok' if pct_prov>=90 else ('warn' if pct_prov>=50 else 'error'),
                        'valor':f'{n_prov}/{n_total} ({pct_prov:.0f}%)','sugerencia':'Botón Normalizar para backfill' if pct_prov<90 else ''})
        items.append({'categoria':'mee','nombre':'MEE con precio_unit','status':'ok' if pct_precio>=70 else ('warn' if pct_precio>=30 else 'error'),
                        'valor':f'{n_precio}/{n_total} ({pct_precio:.0f}%)','sugerencia':'Catalina llena precio cuando recibe SCs'})
    except Exception:
        pass

    # 6) SKUs activos con mappings MEE
    try:
        n_skus = c.execute("""
            SELECT COUNT(*) FROM sku_planeacion_config
            WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado','sin_ventas')
        """).fetchone()[0]
        n_mapeados = c.execute("SELECT COUNT(DISTINCT sku_codigo) FROM sku_mee_config WHERE aplica=1").fetchone()[0]
        pct = (n_mapeados/max(n_skus,1))*100
        items.append({'categoria':'mee','nombre':'SKUs mapeados a MEE','status':'ok' if pct>=80 else ('warn' if pct>=30 else 'warn'),
                        'valor':f'{n_mapeados}/{n_skus} ({pct:.0f}%)','sugerencia':'Auto-mapeo + IA aprende cuando Catalina asigna' if pct<80 else ''})
    except Exception:
        pass

    # 7) Áreas sucias sin limpieza
    try:
        rows = c.execute("""
            SELECT a.codigo, a.nombre FROM areas_planta a
            WHERE a.activo=1 AND a.estado='sucia'
              AND NOT EXISTS (
                SELECT 1 FROM limpieza_profunda_calendario l
                WHERE l.area_codigo = a.codigo
                  AND l.estado IN ('pendiente','asignada','en_proceso')
                  AND date(l.fecha) >= date('now')
              )
        """).fetchall()
        if rows:
            items.append({'categoria':'salas','nombre':'Áreas sucias sin limpieza programada','status':'warn',
                            'valor':f'{len(rows)} áreas','sugerencia':'Click "Auto-asignar pendientes" o el cron 06:30 lo asigna mañana'})
        else:
            items.append({'categoria':'salas','nombre':'Áreas sucias','status':'ok','valor':'todas atendidas','sugerencia':''})
    except Exception:
        pass

    # 8) Producciones próximas sin asignar
    try:
        fecha_max = (fecha_hoy + timedelta(days=7)).isoformat()
        n_pend = c.execute("""
            SELECT COUNT(*) FROM produccion_programada
            WHERE date(fecha_programada) BETWEEN ? AND ?
              AND COALESCE(estado, 'programado') NOT IN ('completado', 'cancelado')
              AND (area_id IS NULL OR (operario_dispensacion_id IS NULL
                   AND operario_elaboracion_id IS NULL
                   AND operario_envasado_id IS NULL))
        """, (fecha_hoy.isoformat(), fecha_max)).fetchone()[0]
        items.append({'categoria':'produccion','nombre':'Producciones próximas sin asignar','status':'ok' if n_pend==0 else 'warn',
                        'valor':str(n_pend),'sugerencia':'Cron 06:30 asigna · o click manual' if n_pend>0 else ''})
    except Exception:
        pass

    # Resumen
    n_ok = sum(1 for it in items if it['status']=='ok')
    n_warn = sum(1 for it in items if it['status']=='warn')
    n_err = sum(1 for it in items if it['status']=='error')

    overall = 'ok' if n_err==0 and n_warn<=2 else ('warn' if n_err==0 else 'error')
    return jsonify({
        'fecha': fecha_hoy.isoformat(),
        'overall_status': overall,
        'kpis': {'ok': n_ok, 'warn': n_warn, 'error': n_err, 'total': len(items)},
        'items': items,
    })


@bp.route('/api/planta/self-heal', methods=['POST'])
def planta_self_heal():
    """Self-healing: arregla problemas comunes detectados.
    Sebastián 1-may-2026: "que todo se programe solo automatico".

    Acciones:
      1. Habilita auto_plan_cron si está deshabilitado
      2. Crea limpieza para áreas sucias sin limpieza programada
      3. Auto-asigna producciones próximos 7d sin área/operarios
      4. Limpia logs viejos (cron_jobs_runs > 30d, auto_plan_runs > 90d)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', 'manual')
    conn = get_db(); c = conn.cursor()
    acciones = []

    # 1) Habilitar cron
    try:
        r = c.execute("SELECT habilitado FROM auto_plan_cron_state WHERE id=1").fetchone()
        if r and not r[0]:
            c.execute("UPDATE auto_plan_cron_state SET habilitado=1, notas='Self-heal habilitado por '||?, activado_por=?, activado_at=datetime('now') WHERE id=1",
                      (user, user))
            acciones.append('Auto-plan cron HABILITADO')
    except Exception as e:
        acciones.append(f'⚠ Error habilitar cron: {e}')

    # 1b) Reset salas con estado stale: 'ocupada' pero sin producción activa
    try:
        rows = c.execute("""
            SELECT a.id, a.codigo FROM areas_planta a
            WHERE a.activo = 1 AND a.estado = 'ocupada'
              AND NOT EXISTS (
                SELECT 1 FROM produccion_programada pp
                WHERE pp.area_id = a.id
                  AND pp.estado IN ('programado','en_proceso','iniciado')
                  AND date(pp.fecha_programada) >= date('now', '-3 days')
              )
        """).fetchall()
        for area_id, codigo in rows:
            c.execute("UPDATE areas_planta SET estado='libre' WHERE id=?", (area_id,))
            acciones.append(f'Reset {codigo}: ocupada→libre (sin producción activa)')
    except Exception as e:
        acciones.append(f'⚠ Error reset salas: {e}')

    # 2) Crear limpieza para sucias sin programación
    try:
        from blueprints.programacion import _crear_limpieza_post_produccion
        rows = c.execute("""
            SELECT a.id, a.codigo, a.nombre FROM areas_planta a
            WHERE a.activo=1 AND a.estado='sucia'
              AND NOT EXISTS (
                SELECT 1 FROM limpieza_profunda_calendario l
                WHERE l.area_codigo = a.codigo
                  AND l.estado IN ('pendiente','asignada','en_proceso')
                  AND date(l.fecha) >= date('now')
              )
        """).fetchall()
        from datetime import date as _d
        for area_id, area_cod, area_nom in rows:
            limp = _crear_limpieza_post_produccion(c, area_id, area_cod, _d.today().isoformat(),
                                                    'self-heal', '', user)
            if limp:
                acciones.append(f'Limpieza creada para {area_cod} (#{limp})')
    except Exception as e:
        acciones.append(f'⚠ Error crear limpiezas: {e}')

    # 3) Auto-asignar producciones pendientes
    try:
        from blueprints.programacion import _auto_asignar_produccion
        from datetime import datetime as _dt, timedelta as _td
        fecha_hoy = _dt.now().date()
        fecha_max = fecha_hoy + _td(days=7)
        rows = c.execute("""
            SELECT id FROM produccion_programada
            WHERE date(fecha_programada) BETWEEN ? AND ?
              AND COALESCE(estado, 'programado') NOT IN ('completado', 'cancelado')
              AND (area_id IS NULL OR (operario_dispensacion_id IS NULL
                   AND operario_elaboracion_id IS NULL
                   AND operario_envasado_id IS NULL))
        """, (fecha_hoy.isoformat(), fecha_max.isoformat())).fetchall()
        n_asign = 0
        for (pid,) in rows:
            res = _auto_asignar_produccion(c, pid, f'self-heal-{user}')
            if res.get('ok'):
                n_asign += 1
        if n_asign:
            acciones.append(f'{n_asign} producciones auto-asignadas')
    except Exception as e:
        acciones.append(f'⚠ Error auto-asignar: {e}')

    # 4) Cleanup logs viejos
    try:
        n_runs = c.execute("DELETE FROM cron_jobs_runs WHERE date(ejecutado_at) < date('now', '-30 days')").rowcount
        n_apr = c.execute("DELETE FROM auto_plan_runs WHERE date(ejecutado_at) < date('now', '-90 days')").rowcount
        if n_runs or n_apr:
            acciones.append(f'Limpieza logs: {n_runs} cron_jobs_runs + {n_apr} auto_plan_runs')
    except Exception as e:
        acciones.append(f'⚠ Error cleanup logs: {e}')

    conn.commit()
    return jsonify({
        'ok': True,
        'acciones': acciones,
        'total': len(acciones),
        'mensaje': f'✅ Self-heal completado · {len(acciones)} acciones' if acciones else '✅ Sistema en buen estado · nada que reparar',
    })


@bp.route('/api/planta/diagnostico-calendar', methods=['GET'])
def diagnostico_calendar():
    """Diagnóstico production-grade del Calendar feed.

    Sebastián 1-may-2026: 'haz que el calendario funcione, ya está conectado'.
    Devuelve TODO lo necesario para diagnosticar por qué no jala eventos:
      · Estado env vars (GCAL_ICAL_URL / GOOGLE_API_KEY / CALENDAR_ID)
      · URL siendo usada (oculta partes secretas)
      · Resultado fetch en vivo (force_refresh=True, sin cache)
      · Status HTTP / encoding / longitud feed / error específico
      · Primeros 10 eventos parseados con título + fecha
      · Sample matching producto-evento (top 3 productos activos)
      · Tiempo de fetch en ms
      · Sugerencias específicas por caso
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    import os as _os
    import time as _time
    import urllib.request as _ur

    gcal_url = _os.environ.get('GCAL_ICAL_URL', '').strip()
    google_key = _os.environ.get('GOOGLE_API_KEY', '').strip()
    cal_id = _os.environ.get('CALENDAR_ID', '').strip()

    diag = {
        'fecha_diagnostico': datetime.now().isoformat(),
        'env_vars': {
            'GCAL_ICAL_URL_configurado': bool(gcal_url),
            'GCAL_ICAL_URL_preview': (gcal_url[:35] + '…' + gcal_url[-12:]) if len(gcal_url) > 50 else (gcal_url or '(no definido)'),
            'GOOGLE_API_KEY_configurado': bool(google_key),
            'GOOGLE_API_KEY_preview': (google_key[:8] + '…') if google_key else '(no definido)',
            'CALENDAR_ID': cal_id or '(no definido)',
        },
    }

    # Fetch en vivo con force_refresh + cronometrar
    t0 = _time.time()
    try:
        from blueprints.programacion import _fetch_calendar_events
        result = _fetch_calendar_events(days_ahead=60) or {}
        eventos = result.get('events') or []
        ms = int((_time.time() - t0) * 1000)
        diag['fetch'] = {
            'source': result.get('source', 'unknown'),
            'duracion_ms': ms,
            'error': result.get('error'),
            'total_eventos_60d': len(eventos),
        }
        diag['eventos_sample'] = [
            {'fecha': e.get('fecha'), 'titulo': (e.get('titulo') or '')[:80],
             'descripcion_preview': (e.get('descripcion') or '')[:60]}
            for e in eventos[:10]
        ]
    except Exception as e:
        ms = int((_time.time() - t0) * 1000)
        import traceback
        diag['fetch'] = {
            'source': 'fail',
            'duracion_ms': ms,
            'error': str(e),
            'traceback_preview': traceback.format_exc()[-500:],
            'total_eventos_60d': 0,
        }
        diag['eventos_sample'] = []

    # Test directo de URL si hay GCAL_ICAL_URL (extra info para debug)
    # Test directo URL: solo si está configurada (corre siempre, no solo si falla
    # el fetch principal). Permite confirmar visualmente que la URL responde OK.
    if gcal_url:
        try:
            req = _ur.Request(gcal_url, headers={'User-Agent': 'EspagiRIA-Diagnostico/1.0'})
            t1 = _time.time()
            with _ur.urlopen(req, timeout=10) as r:
                content = r.read()
                # Bug fix: Content-Type puede venir vacío o None en algunos
                # endpoints de Google Calendar — validar por BODY no header.
                ct = r.headers.get('Content-Type', '') or ''
                # Strip BOM si lo hay (algunos feeds Google lo incluyen)
                content_clean = content.lstrip(b'\xef\xbb\xbf').lstrip()
                es_ical = content_clean[:50].startswith(b'BEGIN:VCALENDAR') or b'BEGIN:VCALENDAR' in content[:200]
                diag['url_test'] = {
                    'status': r.getcode(),
                    'content_type': ct or '(sin header)',
                    'size_bytes': len(content),
                    'duracion_ms': int((_time.time() - t1) * 1000),
                    'preview_500_chars': content[:500].decode('utf-8', errors='replace'),
                    'es_ical_valido': bool(es_ical),
                    'cantidad_VEVENT': content.count(b'BEGIN:VEVENT'),
                }
        except Exception as e:
            diag['url_test'] = {'error': str(e)}

    # Test matching producto-evento (los primeros 3 productos activos)
    diag['matching_test'] = []
    try:
        conn = get_db(); c = conn.cursor()
        eventos = []
        if diag['fetch'].get('total_eventos_60d', 0) > 0:
            from blueprints.programacion import _fetch_calendar_events
            eventos = (_fetch_calendar_events(days_ahead=60) or {}).get('events') or []
        productos_activos = c.execute("""
            SELECT producto_nombre FROM sku_planeacion_config
            WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
            LIMIT 3
        """).fetchall()
        for (prod,) in productos_activos:
            try:
                alias = _alias_calendar_for(c, prod)
            except Exception:
                alias = ''
            mejores = []
            for ev in eventos[:30]:
                try:
                    score = _match_producto_evento(prod, alias, ev.get('titulo',''),
                                                     ev.get('descripcion',''))
                    if score > 0:
                        mejores.append({'evento': (ev.get('titulo') or '')[:50],
                                          'fecha': ev.get('fecha'), 'score': score})
                except Exception:
                    continue
            mejores.sort(key=lambda x: x['score'], reverse=True)
            diag['matching_test'].append({
                'producto': prod[:40],
                'alias': alias or '(ninguno)',
                'top_matches': mejores[:3],
            })
    except Exception:
        pass

    # Sugerencias inteligentes basadas en el caso detectado
    # Sebastián 1-may-2026: no dar falsos positivos cuando todo OK.
    s = []
    eventos_ok = diag['fetch'].get('total_eventos_60d', 0) > 0
    has_url = bool(gcal_url or google_key)

    if not has_url:
        s.append('🚨 Configura GCAL_ICAL_URL en Render: Calendar Producciones → Configuración → Integrar calendario → URL secreta en formato iCal (.ics)')
    elif eventos_ok:
        # FUNCIONA - solo verificar matching
        if diag['matching_test']:
            sin_match = [t for t in diag['matching_test'] if not t['top_matches']]
            scores_bajos = [t for t in diag['matching_test']
                              if t['top_matches'] and t['top_matches'][0]['score'] < 60]
            if len(sin_match) == len(diag['matching_test']):
                s.append('⚠️ Hay eventos pero NINGÚN producto activo hace match. Configura aliases en sku_planeacion_config.alias_calendar (códigos cortos como TRIAC, LBHA, NPHA del Calendar)')
            elif len(sin_match) + len(scores_bajos) >= len(diag['matching_test']) - 1:
                s.append('⚠️ La mayoría de productos tienen match débil o nulo. Revisa aliases en sku_planeacion_config.')
            else:
                s.append(f'✅ Calendar funciona correctamente · {diag["fetch"]["total_eventos_60d"]} eventos en 60 días · matching OK')
        else:
            s.append(f'✅ Calendar conectado · {diag["fetch"]["total_eventos_60d"]} eventos en 60 días')
    else:
        # Hay env var pero NO eventos
        ut = diag.get('url_test', {})
        if ut.get('error'):
            s.append(f'⚠️ La URL configurada falla: {ut["error"]}. Verifica que esté vigente.')
        elif not ut.get('es_ical_valido', True):
            s.append(f'⚠️ La URL no devuelve formato iCal válido. Content-Type: {ut.get("content_type", "?")}. Usa la URL "iCal" no la URL pública.')
        elif ut.get('cantidad_VEVENT', 0) == 0:
            s.append('⚠️ El feed iCal no tiene eventos. Crea eventos en el Calendar Producciones.')
        elif ut.get('cantidad_VEVENT', 0) > 0:
            s.append(f'⚠️ El feed tiene {ut.get("cantidad_VEVENT")} eventos pero ninguno cae en próximos 60 días (¿están en el pasado o muy en el futuro?).')
        else:
            s.append('⚠️ Calendar configurado pero sin eventos detectados en 60d. Revisa que estés usando el Calendar correcto.')

    diag['sugerencias'] = s

    return jsonify(diag)


@bp.route('/api/planta/normalizar-mee', methods=['POST'])
def normalizar_mee():
    """Normaliza datos MEE en bulk:
      1. Backfill proveedor_principal desde maestro_mee.proveedor (idempotente)
      2. Auto-mapping fuzzy SKU → etiqueta/serigrafía/tampografía/plegadiza
         basado en similitud de palabras (descripción MEE vs nombre SKU)

    Body opcional: {dry_run: bool=true, umbral_score: int=40}
    Si dry_run=true, devuelve propuestas sin escribir.
    Si dry_run=false, aplica los cambios.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    dry_run = bool(d.get('dry_run', True))
    umbral = int(d.get('umbral_score', 40))
    conn = get_db(); c = conn.cursor()

    resultado = {
        'dry_run': dry_run,
        'umbral_score': umbral,
        'backfill_proveedor': {'antes': 0, 'despues': 0, 'cambios': []},
        'auto_mapping': {'sugerencias': [], 'aplicados': 0, 'omitidos': 0},
    }

    # 1) BACKFILL PROVEEDOR
    n_antes = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE COALESCE(proveedor_principal,'') != ''").fetchone()[0]
    resultado['backfill_proveedor']['antes'] = n_antes

    rows = c.execute("""
        SELECT cfg.mee_codigo, m.proveedor, m.descripcion
        FROM mee_lead_time_config cfg
          JOIN maestro_mee m ON m.codigo = cfg.mee_codigo
        WHERE COALESCE(cfg.proveedor_principal, '') = ''
          AND COALESCE(m.proveedor, '') != ''
    """).fetchall()
    for codigo, prov, desc in rows:
        resultado['backfill_proveedor']['cambios'].append({
            'mee_codigo': codigo, 'descripcion': (desc or '')[:40],
            'proveedor_asignado': prov,
        })
        if not dry_run:
            origen_nuevo = 'China' if prov == 'China' else None
            if origen_nuevo:
                c.execute("""
                    UPDATE mee_lead_time_config
                       SET proveedor_principal = ?, origen = 'China',
                           actualizado_en = datetime('now'), actualizado_por = ?
                     WHERE mee_codigo = ? AND COALESCE(proveedor_principal,'') = ''
                """, (prov, user, codigo))
            else:
                c.execute("""
                    UPDATE mee_lead_time_config
                       SET proveedor_principal = ?,
                           actualizado_en = datetime('now'), actualizado_por = ?
                     WHERE mee_codigo = ? AND COALESCE(proveedor_principal,'') = ''
                """, (prov, user, codigo))

    n_despues = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE COALESCE(proveedor_principal,'') != ''").fetchone()[0]
    resultado['backfill_proveedor']['despues'] = n_despues if not dry_run else (n_antes + len(rows))

    # 2) AUTO-MAPPING SKU → MEE (etiquetas, serigrafías, tampografías)
    # Solo tipos donde el nombre del SKU suele aparecer en la descripción.
    skus_activos = c.execute("""
        SELECT producto_nombre FROM sku_planeacion_config
        WHERE activo = 1
          AND COALESCE(estado, 'activo') NOT IN ('descontinuado', 'pausado', 'sin_ventas')
    """).fetchall()
    skus_activos = [r[0] for r in skus_activos]

    # MEEs candidatos por categoría
    mees_etiqueta = c.execute("""
        SELECT codigo, descripcion FROM maestro_mee
        WHERE COALESCE(estado,'Activo')='Activo'
          AND COALESCE(categoria,'') IN ('Etiqueta','Serigrafia')
    """).fetchall()

    # Mappings ya existentes (para no duplicar)
    existentes = set()
    try:
        for r in c.execute("SELECT sku_codigo, mee_codigo FROM sku_mee_config").fetchall():
            existentes.add((r[0], r[1]))
    except Exception:
        pass

    for sku in skus_activos:
        palabras_sku = _normalizar_palabras(sku)
        if not palabras_sku:
            continue
        # Buscar el mejor match por categoría
        mejores = {}  # categoria → (codigo, score, desc)
        for cod, desc in mees_etiqueta:
            palabras_mee = _normalizar_palabras(desc)
            score = _score_fuzzy(palabras_sku, palabras_mee)
            if score < umbral:
                continue
            # Determinar categoría/tipo del MEE
            cat_row = next(((c_, d_) for c_, d_ in mees_etiqueta if c_ == cod), None)
            tipo = 'etiqueta' if cod.upper().startswith(('ETIQ', 'EMP-ETIQ')) else (
                   'serigrafia' if cod.upper().startswith('SERIG') else 'etiqueta')
            actual = mejores.get(tipo)
            if not actual or score > actual[1]:
                mejores[tipo] = (cod, score, desc)

        for tipo, (cod, score, desc) in mejores.items():
            if (sku, cod) in existentes:
                continue
            sug = {
                'sku_codigo': sku,
                'mee_codigo': cod,
                'mee_descripcion': desc,
                'componente_tipo': tipo,
                'score': score,
                'cantidad_por_unidad': 1,
            }
            resultado['auto_mapping']['sugerencias'].append(sug)
            if not dry_run:
                try:
                    c.execute("""
                        INSERT INTO sku_mee_config
                          (sku_codigo, mee_codigo, componente_tipo,
                           cantidad_por_unidad, aplica, notas)
                        VALUES (?, ?, ?, 1, 1, ?)
                    """, (sku, cod, tipo,
                          f'Auto-mapeo (score {score}) por similitud nombre — Sebastián 1-may-2026'))
                    resultado['auto_mapping']['aplicados'] += 1
                    existentes.add((sku, cod))
                except sqlite3.IntegrityError:
                    resultado['auto_mapping']['omitidos'] += 1

    if not dry_run:
        conn.commit()

    resultado['mensaje'] = (
        f"{'DRY RUN: ' if dry_run else ''}"
        f"{len(resultado['backfill_proveedor']['cambios'])} proveedores "
        f"{'a backfillear' if dry_run else 'backfilleados'}, "
        f"{len(resultado['auto_mapping']['sugerencias'])} mappings "
        f"{'sugeridos' if dry_run else 'creados'}"
    )
    return jsonify(resultado)


# ════════════════════════════════════════════════════════════════════════
# CRUD config MEE · proveedor + origen + MOQ + lead time + flags
# ════════════════════════════════════════════════════════════════════════

@bp.route('/api/planta/mee-config', methods=['GET'])
def mee_config_listar():
    """Lista combinada maestro_mee + mee_lead_time_config para que Sebastián
    pueda configurar proveedor/MOQ/origen sin tocar SQL."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT m.codigo, m.descripcion, m.categoria, m.stock_actual, m.stock_minimo,
               m.estado,
               COALESCE(cfg.proveedor_principal, ''),
               COALESCE(cfg.origen, 'Local'),
               COALESCE(cfg.lead_time_dias, 30),
               COALESCE(cfg.moq_unidades, 0),
               COALESCE(cfg.precio_unit, 0),
               COALESCE(cfg.disparo_d20, 0),
               COALESCE(cfg.disparo_post_envasado, 0),
               COALESCE(cfg.aplica, 1),
               COALESCE(cfg.notas, ''),
               COALESCE(cfg.actualizado_en, ''),
               COALESCE(cfg.actualizado_por, '')
        FROM maestro_mee m
          LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = m.codigo
        ORDER BY
          CASE COALESCE(m.categoria,'') WHEN 'Envase' THEN 1 WHEN 'Frasco' THEN 2
            WHEN 'Tapa' THEN 3 WHEN 'Gotero' THEN 4 WHEN 'Etiqueta' THEN 5
            WHEN 'Serigrafia' THEN 6 WHEN 'Plegable' THEN 7 ELSE 8 END,
          m.codigo
    """).fetchall()
    cols = ['codigo','descripcion','categoria','stock_actual','stock_minimo','estado',
            'proveedor_principal','origen','lead_time_dias','moq_unidades','precio_unit',
            'disparo_d20','disparo_post_envasado','aplica','notas',
            'actualizado_en','actualizado_por']
    return jsonify({'mees': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/planta/mee-config/<path:codigo>', methods=['PUT'])
def mee_config_actualizar(codigo):
    """Actualiza configuración de un MEE. Body: cualquier subset de campos."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    conn = get_db(); c = conn.cursor()

    # Verificar que existe en maestro_mee
    if not c.execute("SELECT 1 FROM maestro_mee WHERE codigo=?", (codigo,)).fetchone():
        return jsonify({'error': 'MEE no existe en maestro_mee'}), 404

    # Validaciones
    origen = (d.get('origen') or '').strip()
    if origen and origen not in ('China', 'Local', 'Mixto'):
        return jsonify({'error': "origen debe ser China, Local o Mixto"}), 400

    # UPSERT: si no hay fila, crear; si hay, actualizar
    existing = c.execute("SELECT 1 FROM mee_lead_time_config WHERE mee_codigo=?", (codigo,)).fetchone()
    if not existing:
        c.execute("""
            INSERT INTO mee_lead_time_config
              (mee_codigo, proveedor_principal, origen, lead_time_dias,
               moq_unidades, precio_unit, disparo_d20, disparo_post_envasado,
               aplica, notas, actualizado_en, actualizado_por)
            VALUES (?, '', 'Local', 30, 0, 0, 0, 0, 1, '', datetime('now'), ?)
        """, (codigo, user))

    # UPDATE de campos provistos
    sets = []
    params = []
    field_map = {
        'proveedor_principal': str,
        'origen': str,
        'lead_time_dias': int,
        'moq_unidades': int,
        'precio_unit': float,
        'disparo_d20': lambda x: 1 if x else 0,
        'disparo_post_envasado': lambda x: 1 if x else 0,
        'aplica': lambda x: 1 if x else 0,
        'notas': str,
    }
    for field, conv in field_map.items():
        if field in d and d[field] is not None:
            try:
                sets.append(f"{field} = ?")
                params.append(conv(d[field]))
            except (ValueError, TypeError):
                return jsonify({'error': f'{field}: valor inválido'}), 400
    if not sets:
        return jsonify({'ok': True, 'mensaje': 'Sin cambios'})
    sets.append("actualizado_en = datetime('now')")
    sets.append("actualizado_por = ?")
    params.append(user)
    params.append(codigo)
    c.execute(f"UPDATE mee_lead_time_config SET {', '.join(sets)} WHERE mee_codigo = ?", params)
    conn.commit()
    return jsonify({'ok': True, 'codigo': codigo, 'campos_actualizados': len(sets)-2})


@bp.route('/api/planta/sku-mee-config', methods=['GET'])
def sku_mee_config_listar():
    """Lista mappings SKU → componente MEE con detalle del MEE."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    sku_filtro = (request.args.get('sku') or '').strip()
    conn = get_db(); c = conn.cursor()
    sql = """
        SELECT s.id, s.sku_codigo, s.mee_codigo, s.componente_tipo,
               s.cantidad_por_unidad, s.aplica, COALESCE(s.notas, ''),
               m.descripcion, m.categoria, m.stock_actual,
               COALESCE(cfg.proveedor_principal, ''), COALESCE(cfg.origen, '')
        FROM sku_mee_config s
          LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
          LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = s.mee_codigo
    """
    params = ()
    if sku_filtro:
        sql += " WHERE UPPER(TRIM(s.sku_codigo)) = UPPER(TRIM(?))"
        params = (sku_filtro,)
    sql += " ORDER BY s.sku_codigo, s.componente_tipo, s.mee_codigo"
    rows = c.execute(sql, params).fetchall()
    cols = ['id','sku_codigo','mee_codigo','componente_tipo','cantidad_por_unidad',
            'aplica','notas','mee_descripcion','mee_categoria','stock_actual',
            'proveedor','origen']
    return jsonify({'mappings': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/planta/sku-mee-config', methods=['POST'])
def sku_mee_config_crear():
    """Crea un mapping SKU → MEE.
    Body: {sku_codigo, mee_codigo, componente_tipo, cantidad_por_unidad?, aplica?, notas?}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    sku = (d.get('sku_codigo') or '').strip()
    mee = (d.get('mee_codigo') or '').strip()
    comp = (d.get('componente_tipo') or '').strip()
    if not sku or not mee or not comp:
        return jsonify({'error': 'sku_codigo, mee_codigo, componente_tipo requeridos'}), 400
    if comp not in ('envase','tapa','etiqueta','caja','serigrafia','tampografia','plegadiza','otro'):
        return jsonify({'error': 'componente_tipo inválido'}), 400
    cant = float(d.get('cantidad_por_unidad', 1))
    aplica = 1 if d.get('aplica', True) else 0
    notas = (d.get('notas') or '').strip()
    conn = get_db(); c = conn.cursor()
    if not c.execute("SELECT 1 FROM maestro_mee WHERE codigo=?", (mee,)).fetchone():
        return jsonify({'error': f'MEE {mee} no existe en maestro_mee'}), 400
    try:
        c.execute("""
            INSERT INTO sku_mee_config
              (sku_codigo, mee_codigo, componente_tipo, cantidad_por_unidad, aplica, notas)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (sku, mee, comp, cant, aplica, notas))
        conn.commit()
        return jsonify({'ok': True, 'id': c.lastrowid})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Ya existe ese mapping (sku, mee)'}), 409


@bp.route('/api/planta/sku-mee-config/<int:mid>', methods=['PUT', 'DELETE'])
def sku_mee_config_modificar(mid):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'DELETE':
        c.execute("DELETE FROM sku_mee_config WHERE id=?", (mid,))
        conn.commit()
        return jsonify({'ok': True, 'eliminado': mid})
    d = request.json or {}
    sets = []
    params = []
    for f, conv in [('cantidad_por_unidad', float),
                     ('aplica', lambda x: 1 if x else 0),
                     ('componente_tipo', str), ('notas', str)]:
        if f in d and d[f] is not None:
            sets.append(f"{f} = ?"); params.append(conv(d[f]))
    if not sets:
        return jsonify({'ok': True, 'mensaje': 'Sin cambios'})
    params.append(mid)
    c.execute(f"UPDATE sku_mee_config SET {', '.join(sets)} WHERE id = ?", params)
    conn.commit()
    return jsonify({'ok': True, 'id': mid})


@bp.route('/api/planta/auto-sc-mee-status', methods=['GET'])
def auto_sc_mee_status():
    """Status del Auto-SC MEE para el panel del dashboard."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from datetime import date, timedelta
    conn = get_db(); c = conn.cursor()
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)

    def _last_run(tipo):
        try:
            row = c.execute("""
                SELECT ejecutado_at, compras_creadas, payload_json, error, emails_enviados
                FROM auto_plan_runs WHERE tipo = ?
                ORDER BY id DESC LIMIT 1
            """, (tipo,)).fetchone()
            if not row: return None
            return {
                'ejecutado_at': row[0], 'scs_creadas': row[1] or 0,
                'payload_json': row[2] or '', 'error': row[3] or '',
                'emails_enviados': row[4] or 0,
            }
        except Exception:
            return None

    # Contar configuración MEE
    mee_configurados = 0
    mee_con_proveedor = 0
    skus_con_mee = 0
    try:
        mee_configurados = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE aplica=1").fetchone()[0]
        mee_con_proveedor = c.execute("SELECT COUNT(*) FROM mee_lead_time_config WHERE aplica=1 AND COALESCE(proveedor_principal,'') != ''").fetchone()[0]
        skus_con_mee = c.execute("SELECT COUNT(DISTINCT sku_codigo) FROM sku_mee_config WHERE aplica=1").fetchone()[0]
    except Exception:
        pass

    # SCs MEE este mes
    scs_mes = 0
    try:
        scs_mes = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE solicitante='auto-plan-ia' AND categoria='Material de Empaque'
              AND date(fecha) >= ?
        """, (inicio_mes.isoformat(),)).fetchone()[0]
    except Exception:
        pass

    return jsonify({
        'hoy': hoy.isoformat(),
        'last_mensual': _last_run('auto_sc_mee_mensual'),
        'last_urgente': _last_run('auto_sc_mee_urgente'),
        'mee_configurados': mee_configurados,
        'mee_con_proveedor': mee_con_proveedor,
        'skus_con_mee': skus_con_mee,
        'scs_mes_actual': scs_mes,
        'configuracion_lista': mee_con_proveedor > 0 and skus_con_mee > 0,
    })


# ════════════════════════════════════════════════════════════════════════
# MP ROLLING FORECAST — consumo MP acumulado a lo largo del horizonte
# ════════════════════════════════════════════════════════════════════════
# Sebastián (30-abr-2026): "el lunes hay unos productos, ellos pueden usar
# materias primas que usa el del martes ya se lo van a gastar entonces
# debe ir sumando para dar una realidad".

@bp.route('/api/planta/mp-rolling-forecast', methods=['GET'])
def mp_rolling_forecast():
    """Para cada producción futura del Calendar, simula el consumo de MP
    día a día y detecta cuándo cada MP se va a quedar sin stock.

    Query params:
      dias: horizonte en días (default 60)

    Devuelve:
      {
        horizonte_dias: 60,
        producciones: [{fecha, producto, kg, num_mps_consume}],
        materias: [{
          material_id, material_nombre, stock_inicial_g,
          consumo_total_g, saldo_final_g, dias_hasta_stockout,
          fecha_stockout, urgencia,
          consumos: [{fecha, producto, kg_lote, g_consumido, saldo_post}],
          comprar_antes_de, comprar_g_recomendado
        }],
        kpis: {total_producciones, mps_afectadas, mps_stockout, mps_criticas}
      }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    dias_horizonte = int(request.args.get('dias') or 60)
    fecha_hoy = datetime.now().date()
    fecha_limite = fecha_hoy + timedelta(days=dias_horizonte)

    conn = get_db(); c = conn.cursor()

    # Stock map robusto
    try:
        from blueprints.programacion import _get_mp_stock, _norm_mp_name
        mp_stock_map = _get_mp_stock(conn)
    except Exception:
        mp_stock_map = {}
        _norm_mp_name = lambda x: str(x or '').upper().strip()

    # Cargar fórmulas activas
    skus = c.execute("""
        SELECT spc.producto_nombre, fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
          AND COALESCE(spc.estado, 'activo') NOT IN ('descontinuado', 'pausado')
    """).fetchall()
    sku_lote_default = {p: l for p, l in skus}

    # Cache fórmulas
    formulas_cache = {}
    def _get_formula(prod):
        if prod in formulas_cache:
            return formulas_cache[prod]
        rows = c.execute("""
            SELECT material_id, material_nombre, porcentaje
            FROM formula_items
            WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
        """, (prod,)).fetchall()
        formulas_cache[prod] = rows
        return rows

    # Eventos del Calendar dentro del horizonte
    eventos = _calendar_events_cached()
    producciones_planeadas = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if f < fecha_hoy or f > fecha_limite:
            continue
        # Match a un SKU activo
        producto_match = None
        for prod_nom in sku_lote_default.keys():
            alias = _alias_calendar_for(c, prod_nom)
            score = _match_producto_evento(prod_nom, alias, ev.get('titulo'), ev.get('descripcion', ''))
            if score >= 60:
                producto_match = prod_nom
                break
        if not producto_match:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', '')) or sku_lote_default.get(producto_match) or 30
        producciones_planeadas.append({
            'fecha': f,
            'producto': producto_match,
            'kg': kg,
        })
    producciones_planeadas.sort(key=lambda p: p['fecha'])

    # Función para resolver stock por (id, nombre)
    def _stock_de(mat_id, mat_nombre):
        mid = str(mat_id or '').strip()
        if mid in mp_stock_map: return float(mp_stock_map[mid])
        if mid.upper() in mp_stock_map: return float(mp_stock_map[mid.upper()])
        nom_up = str(mat_nombre or '').strip().upper()
        if nom_up and nom_up in mp_stock_map: return float(mp_stock_map[nom_up])
        try:
            nom_norm = _norm_mp_name(mat_nombre or '')
            if nom_norm and nom_norm in mp_stock_map: return float(mp_stock_map[nom_norm])
        except Exception:
            pass
        return 0

    # Acumular consumo por MP
    # mp_data[(material_id, material_nombre)] = {stock_inicial, consumos: [...]}
    mp_data = {}

    for prod_evento in producciones_planeadas:
        items = _get_formula(prod_evento['producto'])
        kg = prod_evento['kg']
        total_g_lote = kg * 1000
        for mat_id, mat_nom, pct in items:
            try:
                pct_f = float(pct or 0)
            except Exception:
                pct_f = 0
            req_g = total_g_lote * pct_f / 100
            if req_g <= 0: continue
            key = (mat_id, mat_nom or mat_id)
            if key not in mp_data:
                mp_data[key] = {
                    'material_id': mat_id,
                    'material_nombre': mat_nom or mat_id,
                    'stock_inicial_g': _stock_de(mat_id, mat_nom),
                    'consumos': [],
                }
            mp_data[key]['consumos'].append({
                'fecha': prod_evento['fecha'].isoformat(),
                'producto': prod_evento['producto'],
                'kg_lote': kg,
                'g_consumido': round(req_g, 1),
            })

    # Calcular saldos rolling y stockout por MP
    materias_out = []
    mps_stockout = 0
    mps_criticas = 0  # stockout antes de 30d

    for key, info in mp_data.items():
        stock = info['stock_inicial_g']
        saldo = stock
        consumo_total = 0
        fecha_stockout = None
        for cons in info['consumos']:
            saldo -= cons['g_consumido']
            consumo_total += cons['g_consumido']
            cons['saldo_post_g'] = round(saldo, 1)
            if fecha_stockout is None and saldo < 0:
                fecha_stockout = cons['fecha']

        dias_hasta_stockout = None
        if fecha_stockout:
            try:
                fs = datetime.strptime(fecha_stockout, '%Y-%m-%d').date()
                dias_hasta_stockout = (fs - fecha_hoy).days
            except Exception:
                pass

        # Urgencia: cuándo hay que comprar
        # Lead time + buffer (si está en mp_lead_time_config)
        lt_row = c.execute("""
            SELECT lead_time_dias, origen, proveedor_principal
            FROM mp_lead_time_config WHERE material_id=?
        """, (info['material_id'],)).fetchone()
        lead = lt_row[0] if lt_row else 14
        origen = lt_row[1] if lt_row else 'local'
        proveedor_sugerido = lt_row[2] if lt_row else ''

        comprar_antes = None
        comprar_g = 0
        if fecha_stockout and dias_hasta_stockout is not None:
            comprar_antes_d = max(0, dias_hasta_stockout - lead - 7)  # +7 buffer
            comprar_antes = (fecha_hoy + timedelta(days=comprar_antes_d)).isoformat()
            # Cuánto: lo que falta + 30d de cobertura adicional
            falta_g = abs(saldo)  # saldo final negativo
            comprar_g = round(falta_g + (consumo_total / max(dias_horizonte, 1)) * 30, 1)

        if fecha_stockout:
            mps_stockout += 1
            if dias_hasta_stockout is not None and dias_hasta_stockout <= 30:
                mps_criticas += 1
        urgencia = 'critica' if dias_hasta_stockout is not None and dias_hasta_stockout <= 15 else \
                   'alta' if dias_hasta_stockout is not None and dias_hasta_stockout <= 30 else \
                   'media' if fecha_stockout else 'ok'

        materias_out.append({
            'material_id': info['material_id'],
            'material_nombre': info['material_nombre'],
            'stock_inicial_g': round(stock, 1),
            'consumo_total_g': round(consumo_total, 1),
            'saldo_final_g': round(saldo, 1),
            'fecha_stockout': fecha_stockout,
            'dias_hasta_stockout': dias_hasta_stockout,
            'lead_time_dias': lead,
            'origen': origen,
            'proveedor_sugerido': proveedor_sugerido,
            'comprar_antes_de': comprar_antes,
            'comprar_g_recomendado': comprar_g,
            'consumos': info['consumos'],
            'num_lotes_que_la_usan': len(info['consumos']),
            'urgencia': urgencia,
        })

    # Ordenar: stockouts primero (más cercanos arriba), luego por consumo total descendente
    def _ord(m):
        if m['dias_hasta_stockout'] is not None:
            return (0, m['dias_hasta_stockout'])
        return (1, -m['consumo_total_g'])
    materias_out.sort(key=_ord)

    return jsonify({
        'horizonte_dias': dias_horizonte,
        'fecha_inicio': fecha_hoy.isoformat(),
        'fecha_fin': fecha_limite.isoformat(),
        'producciones': [{
            'fecha': p['fecha'].isoformat(),
            'producto': p['producto'],
            'kg': p['kg'],
        } for p in producciones_planeadas],
        'materias': materias_out,
        'kpis': {
            'total_producciones': len(producciones_planeadas),
            'mps_afectadas': len(materias_out),
            'mps_stockout': mps_stockout,
            'mps_criticas_30d': mps_criticas,
        },
        'reglas': [
            f'Horizonte: {dias_horizonte} días',
            'Consumo MP acumulado: cada producción descuenta de stock simulado',
            'Stockout = saldo proyectado < 0 en alguna fecha',
            'Comprar antes de = fecha stockout − lead time − 7d buffer',
            'Cantidad recomendada = déficit + 30d cobertura adicional',
        ],
    })


# ════════════════════════════════════════════════════════════════════════
# AUDITOR SEMANAL — email cada lunes 7AM con plan + alertas críticas
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "auditor diario automático (email 9 AM lunes
# con plan semanal + alertas)". Se dispara desde cron Render externo
# o manual desde la app.

def _generar_html_auditor_semanal(c):
    """Genera el HTML del email auditor con plan próxima semana + alertas."""
    fecha_hoy = datetime.now().date()
    nombres_dia = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']

    # 1. Eventos del Calendar próximos 7 días
    eventos = _calendar_events_cached()
    eventos_semana = []
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if fecha_hoy <= f <= fecha_hoy + timedelta(days=7):
            eventos_semana.append({'fecha': f, 'titulo': ev.get('titulo', ''), 'desc': ev.get('descripcion', '')})
    eventos_semana.sort(key=lambda x: x['fecha'])

    # 2. Alertas críticas (vendes 20%+ rápido o lento)
    alertas_criticas = []
    skus_act = c.execute("""
        SELECT spc.producto_nombre, spc.cadencia_dias, fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1
          AND COALESCE(spc.estado, 'activo') NOT IN ('descontinuado', 'pausado')
    """).fetchall()
    for producto, cad_cfg, lote_kg in skus_act:
        alias = _alias_calendar_for(c, producto)
        evs_fut = []
        for ev in eventos:
            try:
                f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            except Exception:
                continue
            if f < fecha_hoy or (f - fecha_hoy).days > 60:
                continue
            score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
            if score < 60: continue
            kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
            evs_fut.append({'fecha': f, 'kg': kg})
        if not evs_fut: continue
        evs_fut.sort(key=lambda x: x['fecha'])
        kg_prox = evs_fut[0]['kg'] or lote_kg or 30
        cad_real = (evs_fut[1]['fecha'] - evs_fut[0]['fecha']).days if len(evs_fut) >= 2 else (cad_cfg or 60)
        fg = _factor_g_por_unidad(c, producto)
        u_lote = (kg_prox * 1000) / max(fg, 1)
        vel_plan = u_lote / max(cad_real + 20, 30)
        vel_real, ftend = _velocidad_total_producto(c, producto)
        vel_real_p = vel_real * (ftend or 1)
        if vel_real_p < 0.05 or vel_plan < 0.05: continue
        ratio = vel_real_p / vel_plan
        if ratio >= 1.20 or ratio <= 0.65:
            alertas_criticas.append({
                'producto': producto,
                'estado': 'ADELANTAR ' + str(int((ratio-1)*100)) + '%' if ratio >= 1.20 else 'REDUCIR LOTE ' + str(int((1-ratio)*100)) + '% menos',
                'urg': 'red' if ratio >= 1.20 else 'purple',
                'proxima': evs_fut[0]['fecha'].isoformat(),
            })

    # 3. Render HTML
    html_eventos = ''
    por_dia = defaultdict(list)
    for e in eventos_semana:
        por_dia[e['fecha']].append(e)
    for delta in range(8):
        d = fecha_hoy + timedelta(days=delta)
        evs = por_dia.get(d, [])
        if not evs and delta > 0: continue  # solo hoy aunque vacío
        col = '#dc2626' if delta == 0 else '#0f172a'
        html_eventos += f'<div style="border-left:4px solid {col};padding:8px 12px;margin-bottom:6px;background:#f8fafc;border-radius:0 6px 6px 0">'
        html_eventos += f'<b>{nombres_dia[d.weekday()]} {d.strftime("%d-%b")}</b>'
        if not evs:
            html_eventos += '<div style="font-size:12px;color:#94a3b8;margin-top:3px">Sin producciones</div>'
        else:
            for e in evs:
                html_eventos += f'<div style="font-size:12px;color:#475569;margin-top:3px">• {e["titulo"]}</div>'
        html_eventos += '</div>'

    html_alertas = ''
    if alertas_criticas:
        for a in alertas_criticas[:8]:
            colores = {'red': '#dc2626', 'purple': '#a855f7'}
            col = colores.get(a['urg'], '#dc2626')
            html_alertas += f'<div style="background:{col}15;border-left:4px solid {col};padding:8px 12px;margin-bottom:6px;border-radius:0 6px 6px 0"><b>{a["producto"]}</b> · {a["estado"]} · próx {a["proxima"]}</div>'
    else:
        html_alertas = '<div style="background:#ecfdf5;border:1px solid #6ee7b7;padding:10px;border-radius:8px;color:#065f46">✅ Sin alertas críticas — todos los SKUs alineados con su plan</div>'

    html = f'''<!DOCTYPE html>
<html><body style="font-family:-apple-system,Segoe UI,sans-serif;background:#f3f4f6;padding:20px;margin:0">
  <div style="max-width:640px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,.08)">
    <div style="background:linear-gradient(135deg,#0f766e,#0891b2);color:#fff;padding:20px">
      <h2 style="margin:0;font-size:20px">📅 Auditor Semanal Producción</h2>
      <p style="margin:4px 0 0;opacity:.9;font-size:13px">Lunes {fecha_hoy.strftime("%d-%b-%Y")} · Ánimus Lab</p>
    </div>
    <div style="padding:20px">
      <h3 style="color:#0f172a;font-size:15px;margin:0 0 10px">🗓️ Producciones esta semana</h3>
      {html_eventos}
      <h3 style="color:#0f172a;font-size:15px;margin:18px 0 10px">🚨 Alertas críticas</h3>
      {html_alertas}
      <div style="margin-top:18px;padding:12px;background:#f1f5f9;border-radius:8px;font-size:11px;color:#64748b">
        Generado automáticamente por EOS Inventarios.<br>
        Para cambiar destinatarios: Plan → Configuración → Email destinatarios.
      </div>
    </div>
  </div>
</body></html>'''
    return html


@bp.route('/api/planta/auditor-semanal-enviar', methods=['POST'])
def auditor_semanal_enviar():
    """Genera y envía el email auditor semanal a los destinatarios configurados.

    Acepta también ?clave=XXX para disparo desde cron externo (Render Cron Job).
    """
    # Auth: sesión compras O clave de cron
    clave_cron = request.args.get('clave', '') or (request.json or {}).get('clave', '')
    clave_esperada = os.environ.get('AUTO_PLAN_CRON_KEY', '')
    es_cron = bool(clave_esperada and clave_cron == clave_esperada)
    if not es_cron and 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    conn = get_db(); c = conn.cursor()

    # Destinatarios
    rows = c.execute("SELECT email FROM email_destinatarios WHERE activo = 1").fetchall()
    destinatarios = [r[0] for r in rows if r[0]]
    if not destinatarios:
        return jsonify({'error': 'Sin destinatarios configurados. Plan → Configuración → Email'}), 400

    # Generar HTML
    html = _generar_html_auditor_semanal(c)
    fecha_hoy = datetime.now().date()
    asunto = f'📅 Auditor Semanal Producción · Sem {fecha_hoy.strftime("%d-%b")}'

    # Enviar
    try:
        import threading, sys, os as _os
        sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
        from notificaciones import SistemaNotificaciones
        notif = SistemaNotificaciones()
        threading.Thread(
            target=notif._enviar_email,
            args=(asunto, html, destinatarios),
            daemon=True
        ).start()
        # Log
        c.execute("""
            INSERT INTO auto_plan_runs (tipo, usuario, productos_planeados, errores, notas, ejecutado_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, ('auditor_semanal', 'cron' if es_cron else session.get('compras_user', '?'),
              0, 0, f'Email enviado a {len(destinatarios)}', datetime.now().isoformat()))
        conn.commit()
        return jsonify({'ok': True, 'destinatarios': len(destinatarios), 'mensaje': f'Email enviado a {len(destinatarios)} destinatarios'})
    except Exception as e:
        log.exception(f'Auditor semanal email error')
        return jsonify({'error': str(e)}), 500


@bp.route('/api/planta/auditor-semanal-preview', methods=['GET'])
def auditor_semanal_preview():
    """Devuelve el HTML del auditor sin enviar (para previsualizar)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    html = _generar_html_auditor_semanal(c)
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ════════════════════════════════════════════════════════════════════════
# Email test real (envía email de prueba al rol/email)
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/auto-plan/configs/emails/test', methods=['POST'])
def email_test():
    u, err, code = _auth()
    if err:
        return err, code
    d = request.json or {}
    email_dest = (d.get('email') or '').strip()
    if not email_dest:
        return jsonify({'error': 'email requerido'}), 400
    asunto = '🤖 Test EOS Auto-Plan'
    html = f"""<!DOCTYPE html>
<html><body style="font-family:-apple-system,sans-serif;background:#f3f4f6;padding:20px">
  <div style="max-width:520px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,.08)">
    <div style="background:linear-gradient(135deg,#7c3aed,#dc2626);color:#fff;padding:20px">
      <h2 style="margin:0">✅ Email funcionando</h2>
      <p style="margin:6px 0 0;opacity:.9;font-size:13px">EOS · Auto-Plan Maestro</p>
    </div>
    <div style="padding:20px;color:#1f2937;font-size:14px">
      <p>Hola,</p>
      <p>Este es un email de prueba enviado desde el módulo Auto-Plan de EOS.</p>
      <p>Si recibes esto, los emails automáticos del cron diario funcionarán correctamente.</p>
      <p style="margin-top:18px;font-size:12px;color:#6b7280">Disparado por: <b>{u}</b><br>Fecha: {datetime.now().isoformat()}</p>
    </div>
  </div>
</body></html>"""
    try:
        import threading, sys, os as _os
        sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
        from notificaciones import SistemaNotificaciones
        notif = SistemaNotificaciones()
        threading.Thread(
            target=notif._enviar_email,
            args=(asunto, html, [email_dest]),
            daemon=True
        ).start()
        return jsonify({'ok': True, 'mensaje': f'Email enviado a {email_dest} (puede tardar 30s)'})
    except Exception as e:
        return jsonify({'error': f'Error enviando email: {e}'}), 500


# ════════════════════════════════════════════════════════════════════════
# Conteo cíclico — endpoints
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/conteo-ciclico/calendario', methods=['GET'])
def conteo_calendario():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 30))
    except Exception:
        dias = 30
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT id, fecha, material_id, material_nombre, categoria_abc,
               asignado_a, estado, stock_esperado_g, stock_real_g, diferencia_g, notas
        FROM conteo_ciclico_calendario
        WHERE fecha BETWEEN date('now') AND date('now', '+' || ? || ' days')
        ORDER BY fecha, categoria_abc, material_nombre
    """, (dias,)).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    pendientes = sum(1 for it in items if it['estado'] == 'programado')
    return jsonify({'items': items, 'total': len(items), 'pendientes': pendientes})


@bp.route('/api/conteo-ciclico/<int:item_id>/registrar', methods=['POST'])
def conteo_registrar(item_id):
    u, err, code = _auth()
    if err:
        return err, code
    d = request.json or {}
    stock_real = d.get('stock_real_g')
    if stock_real is None:
        return jsonify({'error': 'stock_real_g requerido'}), 400
    stock_real = float(stock_real)
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT material_id, stock_esperado_g FROM conteo_ciclico_calendario WHERE id=?",
        (item_id,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Conteo no existe'}), 404
    material_id, stock_esperado = row
    # Calcular esperado si no está seteado: sumar movimientos
    if stock_esperado is None:
        r = c.execute("""
            SELECT COALESCE(SUM(
                CASE WHEN tipo IN ('Ingreso','Ajuste','Devolucion') THEN cantidad
                     WHEN tipo IN ('Salida','Consumo') THEN -cantidad
                     ELSE 0 END
            ), 0) FROM movimientos WHERE material_id=?
        """, (material_id,)).fetchone()
        stock_esperado = float(r[0] if r else 0)
    diferencia = stock_real - (stock_esperado or 0)
    pct = abs(diferencia) / max(1, stock_esperado or 1) * 100
    estado = 'con_diferencia' if pct > 5 else 'cerrado'
    c.execute("""
        UPDATE conteo_ciclico_calendario SET
          stock_esperado_g=?, stock_real_g=?, diferencia_g=?,
          estado=?, terminado_at=datetime('now'), terminado_por=?,
          notas=COALESCE(?, notas)
        WHERE id=?
    """, (stock_esperado, stock_real, diferencia, estado, u, d.get('notas'), item_id))
    # Actualizar último conteo en config
    c.execute("""
        INSERT OR REPLACE INTO conteo_ciclico_config
          (material_id, categoria_abc, frecuencia_dias, ultimo_conteo_fecha,
           ultimo_conteo_diferencia, requiere_validacion, actualizado_en)
        SELECT material_id, COALESCE(categoria_abc,'C'), 90, date('now'), ?,
               CASE WHEN ABS(?) > stock_esperado_g * 0.05 THEN 1 ELSE 0 END,
               datetime('now')
        FROM conteo_ciclico_calendario WHERE id=?
    """, (diferencia, diferencia, item_id))
    conn.commit()
    return jsonify({'ok': True, 'diferencia_g': diferencia, 'estado': estado, 'pct_diferencia': round(pct, 2)})


@bp.route('/api/conteo-ciclico/configs', methods=['GET', 'POST'])
def conteo_configs():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        c.execute("""
            INSERT OR REPLACE INTO conteo_ciclico_config
              (material_id, categoria_abc, frecuencia_dias, actualizado_en)
            VALUES (?, ?, ?, datetime('now'))
        """, (d.get('material_id'), d.get('categoria_abc') or 'C',
              int(d.get('frecuencia_dias') or 90)))
        conn.commit()
        return jsonify({'ok': True})
    rows = c.execute(
        "SELECT * FROM conteo_ciclico_config ORDER BY categoria_abc, material_id"
    ).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'configs': [dict(zip(cols, r)) for r in rows]})


# ════════════════════════════════════════════════════════════════════════
# Perfil riesgo
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/auto-plan/configs/perfil-riesgo', methods=['GET', 'POST'])
def perfil_riesgo():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        c.execute("""
            INSERT OR REPLACE INTO producto_perfil_riesgo
              (producto_nombre, tiene_pigmento, color_descripcion, es_acido,
               requiere_asepsia_extra, riesgo_arrastre_pct, notas, actualizado_en)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """, (
            d.get('producto_nombre'),
            1 if d.get('tiene_pigmento') else 0,
            d.get('color_descripcion'),
            1 if d.get('es_acido') else 0,
            1 if d.get('requiere_asepsia_extra') else 0,
            int(d.get('riesgo_arrastre_pct') or 5),
            d.get('notas'),
        ))
        conn.commit()
        return jsonify({'ok': True})
    rows = c.execute(
        "SELECT * FROM producto_perfil_riesgo ORDER BY producto_nombre"
    ).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'perfiles': [dict(zip(cols, r)) for r in rows]})


# ════════════════════════════════════════════════════════════════════════
# Asistente accionable — tools
# ════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════
# FORECAST MULTI-HORIZONTE — 1sem / 1m / 2m / 3m / 6m / 12m
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "ideal que sea semanal, 1 mes, 2 meses 3 meses
# 6 meses 1 año... la app saca todo, necesidades completas, y quedan
# además las asignaciones semanales".
#
# Devuelve para el horizonte solicitado:
#  - producciones_por_mes (count, kg total, lotes)
#  - mp_consumo_mensual (gramos por material)
#  - envases_consumo_mensual (unidades por código MEE)
#  - capacidad_uso_pct por área por mes
#  - compras_urgentes (envases China que tocan comprar YA)
#  - alertas_capacidad
#  - costo_proyectado (estimado si hay precios de MP)

def _producciones_futuras_kg(c, producto):
    """Suma de kg de TODAS las producciones futuras (BD + Calendar) con
    matcher robusto + parser kg agresivo. Zero-error-enterprise.
    """
    total_kg = 0.0
    eventos_count = 0
    # BD
    rows = c.execute("""
        SELECT cantidad_kg FROM produccion_programada
        WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
          AND estado IN ('pendiente','en_proceso')
          AND fecha_programada >= date('now')
    """, (producto,)).fetchall()
    for r in rows:
        if r[0]:
            total_kg += float(r[0])
            eventos_count += 1
    # Calendar con match robusto
    eventos = _calendar_events_cached()
    alias = _alias_calendar_for(c, producto)
    fecha_hoy = datetime.now().date()
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            if f < fecha_hoy:
                continue
        except Exception:
            continue
        score = _match_producto_evento(producto, alias, ev.get('titulo'), ev.get('descripcion', ''))
        if score < 60:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', ''))
        if kg:
            total_kg += kg
            eventos_count += 1
        else:
            # Sin kg detectado pero match alto: usar lote_size_kg del producto
            r = c.execute(
                "SELECT lote_size_kg FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
                (producto,)
            ).fetchone()
            if r and r[0]:
                total_kg += float(r[0])
                eventos_count += 1
    return total_kg, eventos_count


def _factor_g_por_unidad(c, producto):
    """Devuelve el factor g/unidad. Fallback inteligente por categoría."""
    # 1) Presentación default
    r = c.execute("""
        SELECT factor_g_por_unidad, peso_g, volumen_ml FROM producto_presentaciones
        WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?)) AND activo=1
        ORDER BY es_default DESC, id ASC LIMIT 1
    """, (producto,)).fetchone()
    if r:
        if r[0] and r[0] > 0:
            return float(r[0])
        if r[1] and r[1] > 0:
            return float(r[1])
        if r[2] and r[2] > 0:
            return float(r[2])
    # 2) Fallback por categoría
    cat_row = c.execute(
        "SELECT categoria FROM sku_planeacion_config WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    cat = (cat_row[0] if cat_row else '') or ''
    factores_cat = {
        'limpiador': 150.0,        # 150 mL
        'hidratante': 50.0,        # 50 mL airless
        'suero': 30.0,             # 30 mL típico
        'suero_vit_c': 30.0,
        'suero_ah': 30.0,
        'contorno': 12.0,          # 10-15 mL
        'contorno_ojos': 12.0,
        'maxlash': 4.5,
        'blush_balm': 6.0,
        'mascarilla': 50.0,
        'crema_corporal': 200.0,
        'esencia': 100.0,
    }
    if cat in factores_cat:
        return factores_cat[cat]
    # 3) Heurística por nombre del producto
    nombre_upper = (producto or '').upper()
    if 'LIMPIADOR' in nombre_upper:
        return 150.0
    if 'HIDRATANTE' in nombre_upper or 'EMULSION' in nombre_upper:
        return 50.0
    if 'CONTORNO' in nombre_upper:
        return 12.0
    if 'MAXLASH' in nombre_upper:
        return 4.5
    if 'CREMA CORPORAL' in nombre_upper:
        return 200.0
    if 'MASCARILLA' in nombre_upper:
        return 50.0
    return 30.0  # default suero


def _calcular_demanda_suministro(c, producto, dias_horizonte=60):
    """Lógica MRP: demanda proyectada vs suministro disponible.

    Sebastian (30-abr-2026): "sabemos cuanto se vende y cuanto se necesita
    por shopify, en el calendario aparece que se ha producido hasta hoy
    para excluirlos... dice cuantos kilos para que sepas cuanto hacer".

    Returns:
        dict con velocidad, stock, producciones_futuras_kg/unidades,
        demanda, suministro_total, deficit, cubierto_hasta_dias
    """
    velocidad, factor = _velocidad_total_producto(c, producto)
    velocidad_proj = velocidad * factor
    if velocidad_proj <= 0:
        velocidad_proj = 0.5  # fallback conservador
    stock_actual = _stock_actual_pt(c, producto)
    prod_kg, prod_count = _producciones_futuras_kg(c, producto)
    factor_g = _factor_g_por_unidad(c, producto)
    # Convertir kg producidos a unidades equivalentes
    suministro_futuro_unidades = (prod_kg * 1000) / max(factor_g, 1)
    suministro_total = stock_actual + suministro_futuro_unidades
    demanda = velocidad_proj * dias_horizonte
    cubierto_hasta_dias = suministro_total / max(velocidad_proj, 0.01)
    return {
        'velocidad_dia': velocidad_proj,
        'stock_actual_unid': stock_actual,
        'producciones_futuras_kg': round(prod_kg, 1),
        'producciones_futuras_count': prod_count,
        'suministro_futuro_unidades': round(suministro_futuro_unidades, 0),
        'suministro_total_unidades': round(suministro_total, 0),
        'demanda_unidades': round(demanda, 0),
        'deficit_unidades': round(max(0, demanda - suministro_total), 0),
        'cubierto_hasta_dias': round(cubierto_hasta_dias, 1),
        'factor_g': factor_g,
    }


def _generar_proyeccion_lotes(c, horizonte_meses):
    """Genera proyección de lotes UNIFICADA con _calcular_recomendacion_sku.

    Sebastian (30-abr-2026): el motor y las recomendaciones deben dar lo
    mismo. Para cada SKU calcula la recomendación, y si es accionable
    proyecta los lotes hasta el horizonte respetando su cadencia.
    """
    proyeccion = []
    skus = c.execute("""
        SELECT spc.producto_nombre, spc.cadencia_dias, spc.merma_pct, spc.prioridad,
               fh.lote_size_kg
        FROM sku_planeacion_config spc
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(spc.producto_nombre))
        WHERE spc.activo = 1 AND fh.lote_size_kg IS NOT NULL
          AND COALESCE(spc.estado, 'activo') NOT IN ('descontinuado','pausado')
        ORDER BY spc.prioridad ASC, spc.producto_nombre
    """).fetchall()

    fecha_hoy = datetime.now().date()
    fecha_fin = fecha_hoy + timedelta(days=horizonte_meses * 30)
    LOTES_MAX_POR_DIA = 2
    MARGEN_DIAS = 20
    lotes_por_fecha = {}

    def _ajustar_a_lmv_disponible(fecha_base):
        f = fecha_base
        for _ in range(120):
            while f.weekday() not in (0, 2, 4):
                f += timedelta(days=1)
            iso = f.isoformat()
            ya_hay_proj = lotes_por_fecha.get(iso, 0)
            try:
                ya_hay_bd = c.execute(
                    "SELECT COUNT(*) FROM produccion_programada WHERE date(fecha_programada)=? AND estado IN ('pendiente','en_proceso')",
                    (iso,)
                ).fetchone()[0]
            except Exception:
                ya_hay_bd = 0
            if (ya_hay_bd + ya_hay_proj) < LOTES_MAX_POR_DIA:
                return f
            f += timedelta(days=1)
        return f

    for producto, cadencia, merma, prioridad, lote_kg in skus:
        # Usar la lógica unificada del motor de recomendaciones
        rec = _calcular_recomendacion_sku(c, producto, lote_kg, cadencia, merma, prioridad)
        # Solo proyectar si es accionable
        if rec['urgencia'] in ('innecesaria', 'sin_ventas', 'baja_rotacion', 'inactivo'):
            continue
        if not rec.get('fecha_proxima'):
            continue

        # Cadencia efectiva: histórica > configurada > default 60d
        cadencia_efectiva = (rec.get('cadencia_historica_dias')
                             or rec.get('cadencia_configurada')
                             or 60)
        # Lote efectivo: típico histórico > default
        lote_efectivo = rec.get('lote_tipico_kg') or lote_kg or 30

        try:
            siguiente = datetime.strptime(rec['fecha_proxima'], '%Y-%m-%d').date()
        except Exception:
            siguiente = fecha_hoy + timedelta(days=2)
        siguiente = _ajustar_a_lmv_disponible(siguiente)
        n_lote = 0

        while siguiente <= fecha_fin and n_lote < 30:  # safety
            iso = siguiente.isoformat()
            lotes_por_fecha[iso] = lotes_por_fecha.get(iso, 0) + 1
            proyeccion.append({
                'producto': producto,
                'fecha': iso,
                'mes': siguiente.strftime('%Y-%m'),
                'lote_kg': lote_efectivo,
                'kg_con_merma': lote_efectivo * (1 + (merma or 0) / 100.0),
                'velocidad_dia': rec.get('velocidad_dia') or 0,
                'tipo': 'mrp',
                'cadencia_dias': cadencia_efectiva,
                'razon_mrp': rec.get('razon') or '',
                'lote_num': n_lote + 1,
                'urgencia_inicial': rec['urgencia'],
            })
            n_lote += 1
            siguiente = _ajustar_a_lmv_disponible(siguiente + timedelta(days=cadencia_efectiva))

    return sorted(proyeccion, key=lambda x: x['fecha'])


@bp.route('/api/planta/forecast', methods=['GET'])
def planta_forecast():
    """Forecast multi-horizonte. Querystring:
       ?meses=1|2|3|6|12 (default 3)

    Devuelve plan COMPLETO + necesidades agregadas para el horizonte:
    - producciones_proyectadas (lista cronológica)
    - resumen_mensual: por mes, cuántos lotes, kg total
    - mp_consumo_mensual: por material × mes (g + kg)
    - envases_necesarios: por código envase × mes (unidades)
    - capacidad_uso: por área × mes (% capacidad usada)
    - compras_urgentes: lo que hay que pedir HOY por lead time
    - alertas: capacidad excedida, MP en déficit a futuro
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        meses = int(request.args.get('meses', 3))
    except Exception:
        meses = 3
    meses = max(1, min(meses, 12))

    conn = get_db(); c = conn.cursor()
    proyeccion = _generar_proyeccion_lotes(c, meses)

    # Resumen por mes
    resumen_mensual = {}
    for p in proyeccion:
        m = p['mes']
        if m not in resumen_mensual:
            resumen_mensual[m] = {'lotes': 0, 'kg_total': 0, 'productos_distintos': set()}
        resumen_mensual[m]['lotes'] += 1
        resumen_mensual[m]['kg_total'] += p['kg_con_merma']
        resumen_mensual[m]['productos_distintos'].add(p['producto'])
    for m in resumen_mensual:
        resumen_mensual[m]['productos_distintos'] = len(resumen_mensual[m]['productos_distintos'])
        resumen_mensual[m]['kg_total'] = round(resumen_mensual[m]['kg_total'], 1)

    # Consumo MP por mes
    mp_consumo = {}  # mes -> material_id -> {nombre, gramos}
    for p in proyeccion:
        items = c.execute("""
            SELECT material_id, material_nombre, COALESCE(cantidad_g_por_lote,0), porcentaje
            FROM formula_items
            WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))
        """, (p['producto'],)).fetchall()
        for mat_id, mat_nom, cant_lote, pct in items:
            req_g = cant_lote
            if not req_g and pct:
                req_g = (pct / 100.0) * (p['lote_kg'] or 0) * 1000
            req_g = req_g * (1 + (p.get('merma_pct') or 0) / 100.0)
            m = p['mes']
            if m not in mp_consumo:
                mp_consumo[m] = {}
            if mat_id not in mp_consumo[m]:
                mp_consumo[m][mat_id] = {'nombre': mat_nom, 'gramos': 0}
            mp_consumo[m][mat_id]['gramos'] += req_g

    # Envases por mes (necesita producto_presentaciones)
    envases_consumo = {}
    for p in proyeccion:
        # Buscar presentación default y envase
        pres = c.execute("""
            SELECT envase_codigo, factor_g_por_unidad, etiqueta
            FROM producto_presentaciones
            WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND activo=1
            ORDER BY es_default DESC, id ASC
            LIMIT 1
        """, (p['producto'],)).fetchone()
        if not pres or not pres[0]:
            continue
        env_codigo, factor, etiqueta = pres
        if not factor or factor <= 0:
            # Estimar factor desde lote y volumen ml (asumiendo densidad ~1)
            continue
        unidades = (p['kg_con_merma'] * 1000) / factor
        m = p['mes']
        if m not in envases_consumo:
            envases_consumo[m] = {}
        if env_codigo not in envases_consumo[m]:
            envases_consumo[m][env_codigo] = {'etiqueta': etiqueta, 'unidades': 0, 'productos': set()}
        envases_consumo[m][env_codigo]['unidades'] += unidades
        envases_consumo[m][env_codigo]['productos'].add(p['producto'])
    # Convertir set a list para JSON
    for m in envases_consumo:
        for cod in envases_consumo[m]:
            envases_consumo[m][cod]['productos'] = list(envases_consumo[m][cod]['productos'])
            envases_consumo[m][cod]['unidades'] = int(envases_consumo[m][cod]['unidades'])

    # Capacidad por área por mes (cuántos lotes vs días disponibles L/M/V)
    capacidad_uso = {}
    for p in proyeccion:
        # Asignación implícita: por categoría
        # FAB1=lotes >50kg, FAB2=lotes 20-50kg, FAB3=lotes >100kg, FAB_FLOAT=<20kg
        kg = p['lote_kg'] or 0
        if kg >= 100:
            area = 'FAB3'
        elif kg >= 50:
            area = 'FAB1'
        elif kg >= 20:
            area = 'FAB2'
        else:
            area = 'FAB_FLOAT'
        m = p['mes']
        capacidad_uso.setdefault(m, {}).setdefault(area, 0)
        capacidad_uso[m][area] += 1
    # 12 días LMV/mes promedio (4.3 sem × 3 días)
    DIAS_LMV_MES = 13
    alertas_capacidad = []
    for m, areas in capacidad_uso.items():
        for area, lotes in areas.items():
            uso_pct = round((lotes / DIAS_LMV_MES) * 100)
            capacidad_uso[m][area] = {'lotes': lotes, 'uso_pct': uso_pct}
            if uso_pct > 90:
                alertas_capacidad.append({
                    'mes': m, 'area': area, 'lotes': lotes, 'uso_pct': uso_pct,
                    'severidad': 'critica' if uso_pct > 100 else 'alta',
                    'mensaje': f'{area} en {m}: {lotes} lotes vs ~{DIAS_LMV_MES} días disponibles ({uso_pct}%)'
                })

    # Compras urgentes: envases China con lead 180d
    compras_urgentes = []
    fecha_hoy = datetime.now().date()
    for m, envs in envases_consumo.items():
        for env_codigo, info in envs.items():
            try:
                mes_dt = datetime.strptime(m, '%Y-%m').date()
                dias_hasta_mes = (mes_dt - fecha_hoy).days
            except Exception:
                continue
            # Buscar lead time
            lt = c.execute(
                "SELECT lead_time_dias, origen FROM mp_lead_time_config WHERE material_id=?",
                (env_codigo,)
            ).fetchone()
            if not lt:
                lt = (14, 'local')
            lead, origen = lt
            # Si dias_hasta_mes < lead → debes comprar YA
            if dias_hasta_mes < lead:
                compras_urgentes.append({
                    'envase_codigo': env_codigo,
                    'etiqueta': info['etiqueta'],
                    'unidades_requeridas': info['unidades'],
                    'mes_objetivo': m,
                    'lead_time_dias': lead,
                    'origen': origen,
                    'dias_hasta_mes': dias_hasta_mes,
                    'urgencia': 'critica' if origen in ('china', 'usa', 'europa') else 'alta',
                })

    # Costo proyectado (si tenemos precios de MP)
    costo_total_estimado = 0
    try:
        for m, mats in mp_consumo.items():
            for mat_id, info in mats.items():
                pr = c.execute(
                    "SELECT precio_unitario FROM movimientos WHERE material_id=? AND precio_unitario>0 ORDER BY id DESC LIMIT 1",
                    (mat_id,)
                ).fetchone()
                precio_g = (pr[0] / 1000.0) if pr else 0
                costo_total_estimado += info['gramos'] * precio_g
    except Exception:
        pass

    return jsonify({
        'horizonte_meses': meses,
        'fecha_inicio': fecha_hoy.isoformat(),
        'fecha_fin': (fecha_hoy + timedelta(days=meses * 30)).isoformat(),
        'producciones_proyectadas': proyeccion,
        'resumen_mensual': resumen_mensual,
        'mp_consumo_mensual': mp_consumo,
        'envases_consumo_mensual': envases_consumo,
        'capacidad_uso_mensual': capacidad_uso,
        'compras_urgentes': compras_urgentes,
        'alertas_capacidad': alertas_capacidad,
        'costo_total_estimado': round(costo_total_estimado, 0),
        'kpis': _calcular_kpis_horizonte(c, proyeccion, meses, alertas_capacidad, compras_urgentes),
    })


def _calcular_kpis_horizonte(c, proyeccion, meses, alertas_capacidad, compras_urgentes):
    """KPIs unificados que cuentan TODO lo planeado en el horizonte:
    - Calendar real (eventos matcheados con productos)
    - Motor MRP (proyección)
    - BD (produccion_programada)
    Sebastian: que los KPIs reflejen la realidad completa, no solo lo nuevo.
    """
    fecha_hoy = datetime.now().date()
    fecha_fin = fecha_hoy + timedelta(days=meses * 30)

    # 1) Motor MRP (proyección)
    motor_lotes = len(proyeccion)
    motor_kg = sum(p['kg_con_merma'] for p in proyeccion)
    motor_skus = set(p['producto'] for p in proyeccion)

    # 2) BD interna (produccion_programada futura)
    bd_rows = c.execute("""
        SELECT producto, cantidad_kg FROM produccion_programada
        WHERE estado IN ('pendiente','en_proceso')
          AND fecha_programada >= date('now')
          AND fecha_programada <= date(?)
    """, (fecha_fin.isoformat(),)).fetchall()
    bd_lotes = len(bd_rows)
    bd_kg = sum(float(r[1] or 0) for r in bd_rows)
    bd_skus = set((r[0] or '').strip().upper() for r in bd_rows if r[0])

    # 3) Calendar real (eventos matcheados con kg parseados)
    eventos = _calendar_events_cached()
    cal_lotes = 0
    cal_kg = 0
    cal_skus = set()
    productos_cfg = c.execute("""
        SELECT producto_nombre, alias_calendar FROM sku_planeacion_config WHERE activo=1
    """).fetchall()
    for ev in eventos:
        try:
            f = datetime.strptime((ev.get('fecha') or '')[:10], '%Y-%m-%d').date()
            if f < fecha_hoy or f > fecha_fin:
                continue
        except Exception:
            continue
        # Match
        mejor_score = 0
        mejor_producto = None
        for p in productos_cfg:
            score = _match_producto_evento(p[0], p[1], ev.get('titulo'), ev.get('descripcion', ''))
            if score > mejor_score:
                mejor_score = score
                mejor_producto = p[0]
        if mejor_score < 60:
            continue
        kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion', '')) or 30
        cal_lotes += 1
        cal_kg += kg
        cal_skus.add((mejor_producto or '').strip().upper())

    # SKUs totales con plan
    skus_total = motor_skus.union(bd_skus).union(cal_skus)

    return {
        # KPI unificado: todo planeado
        'total_lotes_proyectados': motor_lotes + bd_lotes + cal_lotes,
        'total_kg_proyectados': round(motor_kg + bd_kg + cal_kg, 1),
        'productos_distintos': len(skus_total),
        # Breakdown por origen
        'desglose': {
            'motor_mrp': {'lotes': motor_lotes, 'kg': round(motor_kg, 1), 'skus': len(motor_skus)},
            'bd_interna': {'lotes': bd_lotes, 'kg': round(bd_kg, 1), 'skus': len(bd_skus)},
            'google_calendar': {'lotes': cal_lotes, 'kg': round(cal_kg, 1), 'skus': len(cal_skus)},
        },
        'meses_con_alerta_capacidad': len(set(a['mes'] for a in alertas_capacidad)),
        'compras_urgentes_count': len(compras_urgentes),
    }


# ════════════════════════════════════════════════════════════════════════
# ASIGNACIÓN SEMANAL — Qué se hace en cada área cada día
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/asignacion-semanal', methods=['GET'])
def planta_asignacion_semanal():
    """Vista por área × día de la semana.

    Sebastian (30-abr-2026): "asignaciones semanales que sería diferente
    pues que se hace en cada área".

    Querystring: ?fecha=YYYY-MM-DD (default = lunes de esta semana)

    Devuelve:
      areas: [{codigo, nombre, dias: {lunes, martes, miercoles, jueves, viernes}}]
      cada día tiene: producciones[], envasados[], conteos[], limpiezas[], operarios[]
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    fecha_str = request.args.get('fecha')
    if fecha_str:
        try:
            base = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except Exception:
            base = datetime.now().date()
    else:
        base = datetime.now().date()
    # Encontrar lunes de esa semana
    while base.weekday() != 0:
        base -= timedelta(days=1)
    dias_semana = [base + timedelta(days=i) for i in range(5)]  # L-V
    nombres_dia = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']

    conn = get_db(); c = conn.cursor()
    fecha_inicio = dias_semana[0].isoformat()
    fecha_fin = dias_semana[-1].isoformat()

    # Cargar áreas activas con flag de limpieza profunda
    areas = c.execute("""
        SELECT codigo, nombre, requiere_limpieza_profunda, puede_producir, puede_envasar, tipo
        FROM areas_planta WHERE activo=1
        ORDER BY orden
    """).fetchall()

    # Producciones programadas en la semana
    prods = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
               pp.area_id, ap.codigo as area_codigo,
               op_disp.nombre || ' ' || COALESCE(op_disp.apellido,'') as op_disp_nombre,
               op_elab.nombre || ' ' || COALESCE(op_elab.apellido,'') as op_elab_nombre,
               op_env.nombre  || ' ' || COALESCE(op_env.apellido,'')  as op_env_nombre,
               pp.estado
        FROM produccion_programada pp
        LEFT JOIN areas_planta ap ON ap.id = pp.area_id
        LEFT JOIN operarios_planta op_disp ON op_disp.id = pp.operario_dispensacion_id
        LEFT JOIN operarios_planta op_elab ON op_elab.id = pp.operario_elaboracion_id
        LEFT JOIN operarios_planta op_env  ON op_env.id  = pp.operario_envasado_id
        WHERE pp.fecha_programada BETWEEN ? AND ?
          AND pp.estado IN ('pendiente','en_proceso')
    """, (fecha_inicio, fecha_fin)).fetchall()

    # Limpieza profunda calendario
    limpiezas = c.execute("""
        SELECT fecha, area_codigo, asignado_a, estado, razon_asignacion
        FROM limpieza_profunda_calendario
        WHERE fecha BETWEEN ? AND ?
    """, (fecha_inicio, fecha_fin)).fetchall()

    # Conteos cíclicos
    conteos = c.execute("""
        SELECT fecha, material_id, material_nombre, asignado_a, estado, categoria_abc
        FROM conteo_ciclico_calendario
        WHERE fecha BETWEEN ? AND ?
    """, (fecha_inicio, fecha_fin)).fetchall()

    # Tareas operativas (acondicionamiento, etc)
    tareas = c.execute("""
        SELECT id, titulo, tipo, fecha_objetivo, asignado_a, producto_relacionado, estado
        FROM tareas_operativas
        WHERE fecha_objetivo BETWEEN ? AND ?
          AND estado IN ('pendiente','en_proceso')
    """, (fecha_inicio, fecha_fin)).fetchall()

    # Armar estructura por área
    resultado = []
    for area_codigo, area_nombre, req_limp, puede_prod, puede_env, tipo in areas:
        area_data = {
            'codigo': area_codigo, 'nombre': area_nombre,
            'tipo': tipo, 'requiere_limpieza_profunda': bool(req_limp),
            'puede_producir': bool(puede_prod), 'puede_envasar': bool(puede_env),
            'dias': {},
        }
        for i, dia in enumerate(dias_semana):
            dia_str = dia.isoformat()
            nombre_dia = nombres_dia[i]
            es_lmv = dia.weekday() in (0, 2, 4)
            area_data['dias'][nombre_dia] = {
                'fecha': dia_str,
                'es_dia_produccion': es_lmv,
                'es_dia_acond_conteo': not es_lmv,
                'producciones': [],
                'limpiezas': [],
                'conteos': [],
                'tareas': [],
            }
        # Llenar producciones
        for p in prods:
            if p[5] != area_codigo:
                continue
            try:
                f = datetime.strptime(p[2][:10], '%Y-%m-%d').date()
                idx = (f - dias_semana[0]).days
                if idx < 0 or idx > 4:
                    continue
                nd = nombres_dia[idx]
                area_data['dias'][nd]['producciones'].append({
                    'id': p[0], 'producto': p[1], 'lotes': p[3],
                    'op_dispensacion': (p[6] or '').strip() or None,
                    'op_elaboracion': (p[7] or '').strip() or None,
                    'op_envasado': (p[8] or '').strip() or None,
                    'estado': p[9],
                })
            except Exception:
                pass
        # Llenar limpiezas
        for l in limpiezas:
            if l[1] != area_codigo:
                continue
            try:
                f = datetime.strptime(l[0][:10], '%Y-%m-%d').date()
                idx = (f - dias_semana[0]).days
                if 0 <= idx <= 4:
                    area_data['dias'][nombres_dia[idx]]['limpiezas'].append({
                        'asignado_a': l[2], 'estado': l[3], 'razon': l[4]
                    })
            except Exception:
                pass
        # Conteos cíclicos van solo a áreas de almacén (ALMP, ALMPT, ALM_MP)
        if area_codigo in ('ALMP', 'ALMPT', 'ALM_MP', 'BDG'):
            for cn in conteos:
                try:
                    f = datetime.strptime(cn[0][:10], '%Y-%m-%d').date()
                    idx = (f - dias_semana[0]).days
                    if 0 <= idx <= 4:
                        area_data['dias'][nombres_dia[idx]]['conteos'].append({
                            'material': cn[2] or cn[1],
                            'asignado_a': cn[3], 'estado': cn[4], 'abc': cn[5],
                        })
                except Exception:
                    pass
        resultado.append(area_data)

    # Tareas no asociadas a área: sumarizar aparte
    tareas_globales = []
    for t in tareas:
        try:
            f = datetime.strptime((t[3] or '')[:10], '%Y-%m-%d').date()
            idx = (f - dias_semana[0]).days
            if 0 <= idx <= 4:
                tareas_globales.append({
                    'id': t[0], 'titulo': t[1], 'tipo': t[2],
                    'fecha': nombres_dia[idx], 'asignado_a': t[4],
                    'producto': t[5], 'estado': t[6],
                })
        except Exception:
            pass

    return jsonify({
        'semana_inicio': fecha_inicio,
        'semana_fin': fecha_fin,
        'dias': [{'nombre': nombres_dia[i], 'fecha': dias_semana[i].isoformat(),
                  'weekday': dias_semana[i].weekday(),
                  'es_lmv': dias_semana[i].weekday() in (0, 2, 4)} for i in range(5)],
        'areas': resultado,
        'tareas_globales': tareas_globales,
    })


# ════════════════════════════════════════════════════════════════════════
# Dossier de Lote PDF — INVIMA / trazabilidad completa
# ════════════════════════════════════════════════════════════════════════
@bp.route('/api/planta/dossier-lote/<lote>', methods=['GET'])
def dossier_lote_pdf(lote):
    """Genera PDF con TODA la trazabilidad de un lote PT:
       - Datos producto + presentación
       - Producción programada (fecha, área, operarios)
       - Eventos de envasado
       - Resultados microbiológicos
       - Disposición de cola_liberacion
       - Movimientos de MP consumidas

    Sebastian: "auto-evidencia regulatoria (INVIMA)" — sirve como
    documentación lista para auditoría."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from flask import Response
    conn = get_db(); c = conn.cursor()

    # Datos envasado
    env = c.execute("""
        SELECT pe.id, pe.produccion_id, pe.producto_nombre, pe.lote,
               pe.presentacion_etiqueta, pe.unidades_planeadas,
               pe.unidades_envasadas, pe.envase_codigo, pe.iniciado_at,
               pe.iniciado_por, pe.terminado_at, pe.terminado_por,
               pe.estado, pp.fecha_programada, pp.cantidad_kg,
               ap.codigo as area_codigo, ap.nombre as area_nombre
        FROM produccion_envasado pe
        LEFT JOIN produccion_programada pp ON pp.id = pe.produccion_id
        LEFT JOIN areas_planta ap ON ap.id = pp.area_id
        WHERE pe.lote = ?
        ORDER BY pe.id DESC LIMIT 1
    """, (lote,)).fetchone()

    if not env:
        return jsonify({'error': 'Lote no encontrado en envasados'}), 404

    # Resultados micro
    micro_rows = c.execute("""
        SELECT microorganismo, valor, estado, fecha_analisis, deadline_resultado
        FROM calidad_micro_resultados
        WHERE lote = ? ORDER BY fecha_analisis DESC
    """, (lote,)).fetchall()

    # Cola liberación
    cola = c.execute("""
        SELECT estado, disposicion, fecha_envasado, fecha_min_liberacion,
               aprobado_por, aprobado_at, notas
        FROM cola_liberacion WHERE lote = ?
        ORDER BY id DESC LIMIT 1
    """, (lote,)).fetchone()

    # Operarios asignados a la producción
    ops = c.execute("""
        SELECT op.nombre, op.apellido, op.rol_predeterminado
        FROM produccion_programada pp
        LEFT JOIN operarios_planta op ON op.id IN (
            pp.operario_dispensacion_id, pp.operario_elaboracion_id,
            pp.operario_envasado_id, pp.operario_acondicionamiento_id
        )
        WHERE pp.id = ? AND op.id IS NOT NULL
    """, (env[1],)).fetchall() if env[1] else []

    # Generar PDF
    try:
        from fpdf import FPDF
    except ImportError:
        return jsonify({'error': 'fpdf no disponible'}), 500

    class PDF(FPDF):
        def footer(self):
            self.set_y(-12)
            self.set_font('Helvetica', 'I', 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 6, f'Dossier · Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} · Pag. {self.page_no()}', align='C')

    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=18)
    W = pdf.w - 20

    # Header
    pdf.set_fill_color(124, 58, 237)
    pdf.rect(10, 10, W, 22, 'F')
    pdf.set_xy(12, 13)
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(W - 4, 8, 'ESPAGIRIA LABORATORIO S.A.S.', ln=True)
    pdf.set_x(12)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(220, 220, 220)
    pdf.cell(W - 4, 5, 'Dossier de lote · Trazabilidad INVIMA', ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(6)

    # Título
    pdf.set_font('Helvetica', 'B', 13)
    pdf.cell(W, 8, f'DOSSIER LOTE: {lote}', ln=True, align='C')
    pdf.ln(4)

    # Producto
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(W, 7, ' PRODUCTO', border=1, fill=True, ln=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(50, 6, '  Nombre:', border='LB')
    pdf.cell(W - 50, 6, str(env[2] or ''), border='RB', ln=True)
    pdf.cell(50, 6, '  Presentación:', border='LB')
    pdf.cell(W - 50, 6, str(env[4] or '—'), border='RB', ln=True)
    pdf.cell(50, 6, '  Envase:', border='LB')
    pdf.cell(W - 50, 6, str(env[7] or '—'), border='RB', ln=True)
    pdf.cell(50, 6, '  Cantidad lote:', border='LB')
    pdf.cell(W - 50, 6, f'{env[14] or 0} kg', border='RB', ln=True)
    pdf.cell(50, 6, '  Unidades envasadas:', border='LB')
    pdf.cell(W - 50, 6, str(env[6] or 'pendiente'), border='RB', ln=True)
    pdf.ln(4)

    # Producción
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(W, 7, ' PRODUCCIÓN', border=1, fill=True, ln=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(50, 6, '  Fecha programada:', border='LB')
    pdf.cell(W - 50, 6, str(env[13] or '—'), border='RB', ln=True)
    pdf.cell(50, 6, '  Área:', border='LB')
    pdf.cell(W - 50, 6, f'{env[16] or "—"} ({env[15] or "—"})', border='RB', ln=True)
    if ops:
        pdf.cell(50, 6, '  Operarios:', border='LB')
        ops_txt = ', '.join(f'{o[0]} {o[1] or ""}'.strip() for o in ops)
        pdf.cell(W - 50, 6, ops_txt[:80], border='RB', ln=True)
    pdf.ln(4)

    # Envasado
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(W, 7, ' ENVASADO', border=1, fill=True, ln=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(50, 6, '  Iniciado:', border='LB')
    pdf.cell(W - 50, 6, f'{env[8] or "—"} por {env[9] or "—"}', border='RB', ln=True)
    pdf.cell(50, 6, '  Terminado:', border='LB')
    pdf.cell(W - 50, 6, f'{env[10] or "pendiente"} por {env[11] or "—"}', border='RB', ln=True)
    pdf.cell(50, 6, '  Estado:', border='LB')
    pdf.cell(W - 50, 6, str(env[12] or '—'), border='RB', ln=True)
    pdf.ln(4)

    # Microbiológicos
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(W, 7, ' RESULTADOS MICROBIOLÓGICOS', border=1, fill=True, ln=True)
    if not micro_rows:
        pdf.set_font('Helvetica', 'I', 9)
        pdf.cell(W, 6, '  Sin análisis registrados', border='LRB', ln=True)
    else:
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(250, 250, 250)
        pdf.cell(50, 6, '  Microorganismo', border=1, fill=True)
        pdf.cell(30, 6, 'Valor', border=1, fill=True)
        pdf.cell(40, 6, 'Estado', border=1, fill=True)
        pdf.cell(W - 120, 6, 'Fecha', border=1, fill=True, ln=True)
        pdf.set_font('Helvetica', '', 8)
        for m in micro_rows:
            pdf.cell(50, 5, '  ' + str(m[0] or '')[:30], border='LR')
            pdf.cell(30, 5, str(m[1] or '—'), border='R')
            pdf.cell(40, 5, str(m[2] or '—'), border='R')
            pdf.cell(W - 120, 5, str(m[3] or '—'), border='R', ln=True)
        pdf.cell(W, 0, '', border='T', ln=True)
    pdf.ln(4)

    # Liberación
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(W, 7, ' LIBERACIÓN', border=1, fill=True, ln=True)
    pdf.set_font('Helvetica', '', 9)
    if cola:
        pdf.cell(50, 6, '  Estado:', border='LB')
        pdf.cell(W - 50, 6, str(cola[0] or '—'), border='RB', ln=True)
        pdf.cell(50, 6, '  Disposición:', border='LB')
        pdf.cell(W - 50, 6, str(cola[1] or 'pendiente'), border='RB', ln=True)
        pdf.cell(50, 6, '  Aprobado por:', border='LB')
        pdf.cell(W - 50, 6, f'{cola[4] or "—"} el {cola[5] or "—"}', border='RB', ln=True)
        if cola[6]:
            pdf.cell(50, 6, '  Notas:', border='LB')
            pdf.cell(W - 50, 6, str(cola[6])[:80], border='RB', ln=True)
    else:
        pdf.set_font('Helvetica', 'I', 9)
        pdf.cell(W, 6, '  Sin registro de liberación', border='LRB', ln=True)
    pdf.ln(8)

    # Footer firma
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(W, 5, f'Dossier auto-generado por EOS Auto-Plan · {datetime.now().isoformat()}', align='C', ln=True)

    pdf_bytes = pdf.output(dest='S')
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin-1')
    return Response(
        pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'inline; filename=dossier_{lote}.pdf'}
    )


@bp.route('/api/asistente/tool/<tool_name>', methods=['POST'])
def asistente_tool(tool_name):
    """Endpoints de tools que el asistente Claude puede invocar."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    args = request.json or {}
    conn = get_db(); c = conn.cursor()

    resultado = {}
    exitoso = True
    try:
        if tool_name == 'listar_alertas':
            # Buscar alertas activas (last auto_plan_run + envases déficit)
            r = c.execute("""
                SELECT ejecutado_at, alertas_criticas FROM auto_plan_runs
                ORDER BY id DESC LIMIT 1
            """).fetchone()
            resultado = {
                'ultimo_run': r[0] if r else None,
                'alertas_criticas': r[1] if r else 0,
            }
        elif tool_name == 'listar_producciones_proximas':
            dias = int(args.get('dias', 14))
            rows = c.execute("""
                SELECT producto, fecha_programada, lotes, estado
                FROM produccion_programada
                WHERE fecha_programada >= date('now')
                  AND fecha_programada <= date('now', '+' || ? || ' days')
                  AND estado IN ('pendiente','en_proceso')
                ORDER BY fecha_programada
            """, (dias,)).fetchall()
            resultado = {
                'producciones': [dict(zip(['producto', 'fecha', 'lotes', 'estado'], r)) for r in rows],
                'total': len(rows),
            }
        elif tool_name == 'capacidad_producir_sku':
            producto = args.get('producto', '').strip()
            fh = c.execute(
                "SELECT lote_size_kg FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
                (producto,)
            ).fetchone()
            if not fh:
                resultado = {'error': 'producto sin fórmula'}
            else:
                lote_kg = fh[0] or 0
                resultado = {'producto': producto, 'lote_kg': lote_kg, 'puede': lote_kg > 0}
        elif tool_name == 'ejecutar_auto_plan':
            if user not in ADMIN_USERS:
                exitoso = False
                resultado = {'error': 'Solo admin puede ejecutar auto-plan'}
            else:
                from blueprints.auto_plan_jobs import ejecutar_auto_plan_diario
                from flask import current_app
                import threading
                threading.Thread(
                    target=ejecutar_auto_plan_diario,
                    args=(current_app._get_current_object(),),
                    daemon=True
                ).start()
                resultado = {'mensaje': 'Auto-plan ejecutándose en background'}
        else:
            exitoso = False
            resultado = {'error': f'tool desconocido: {tool_name}'}
    except Exception as e:
        exitoso = False
        resultado = {'error': str(e)}

    # Log la acción
    try:
        c.execute("""
            INSERT INTO asistente_acciones_log
              (usuario, tool_invocado, tool_args, tool_resultado, exitoso)
            VALUES (?, ?, ?, ?, ?)
        """, (user, tool_name, json.dumps(args, ensure_ascii=False)[:500],
              json.dumps(resultado, ensure_ascii=False)[:1000], 1 if exitoso else 0))
        conn.commit()
    except Exception:
        pass

    return jsonify({'ok': exitoso, 'resultado': resultado})


@bp.route('/api/auto-plan/configs/emails', methods=['GET', 'POST'])
def configs_emails():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        rol = (d.get('rol') or '').strip()
        email = (d.get('email') or '').strip()
        if not rol:
            return jsonify({'error': 'rol requerido'}), 400
        c.execute("""
            UPDATE email_destinatarios_config SET
              email=?, nombre=COALESCE(?, nombre),
              recibe_resumen_diario=COALESCE(?, recibe_resumen_diario),
              recibe_alertas_criticas=COALESCE(?, recibe_alertas_criticas),
              recibe_compras_aprob=COALESCE(?, recibe_compras_aprob),
              recibe_calidad=COALESCE(?, recibe_calidad),
              recibe_agenda_personal=COALESCE(?, recibe_agenda_personal),
              actualizado_en=datetime('now')
            WHERE rol=?
        """, (
            email, d.get('nombre'),
            d.get('recibe_resumen_diario'), d.get('recibe_alertas_criticas'),
            d.get('recibe_compras_aprob'), d.get('recibe_calidad'),
            d.get('recibe_agenda_personal'), rol,
        ))
        if c.rowcount == 0:
            # Insertar nuevo si no existe
            c.execute("""
                INSERT INTO email_destinatarios_config
                  (rol, nombre, email, recibe_resumen_diario,
                   recibe_alertas_criticas, recibe_compras_aprob, recibe_calidad,
                   recibe_agenda_personal, activo, actualizado_en)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, datetime('now'))
            """, (
                rol, d.get('nombre'), email,
                d.get('recibe_resumen_diario', 1),
                d.get('recibe_alertas_criticas', 0),
                d.get('recibe_compras_aprob', 0),
                d.get('recibe_calidad', 0),
                d.get('recibe_agenda_personal', 0),
            ))
        conn.commit()
        return jsonify({'ok': True})
    rows = c.execute(
        "SELECT * FROM email_destinatarios_config WHERE activo=1 ORDER BY rol"
    ).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'configs': [dict(zip(cols, r)) for r in rows]})
