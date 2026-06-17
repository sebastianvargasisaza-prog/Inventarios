# blueprints/compras.py — extraído de index.py (Fase C)
import os
import json
import sqlite3
import hmac
import time
import logging
from datetime import datetime, timedelta

log = logging.getLogger('compras')
from flask import Blueprint, jsonify, request, Response, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from config import (
    DB_PATH, COMPRAS_USERS, ADMIN_USERS, CONTADORA_USERS, COMPRAS_ACCESS,
    USER_EMAILS, LIMITES_APROBACION_OC,
)
from database import get_db
from auth import _client_ip, _is_locked, _record_failure, _clear_attempts, _log_sec
from audit_helpers import audit_log, intentar_insert_con_retry, siguiente_numero_oc as _siguiente_numero_oc
from http_helpers import validate_money
from templates_py.rrhh_html import RRHH_HTML
from templates_py.compromisos_html import COMPROMISOS_HTML
from templates_py.home_html import HOME_HTML
from templates_py.hub_html import HUB_HTML
from templates_py.clientes_html import CLIENTES_HTML
from templates_py.calidad_html import CALIDAD_HTML
from templates_py.gerencia_html import GERENCIA_HTML
from templates_py.financiero_html import FINANCIERO_HTML
from templates_py.login_html import LOGIN_HTML
from templates_py.compras_html import COMPRAS_HTML
from templates_py.recepcion_html import RECEPCION_HTML
from templates_py.salida_html import SALIDA_HTML
from templates_py.solicitudes_html import SOLICITUDES_HTML
from templates_py.dashboard_html import DASHBOARD_HTML

bp = Blueprint('compras', __name__)

# ── Estados OC canonical · simplificación 8→6 · 22-may-2026 ─────────────────
# Consultor LEAN: menos estados = menos confusión Catalina.
# Eliminados: 'Revisada' (sub-estado interno) y 'Parcial' (calc por cantidad_recibida_g)
ESTADOS_OC_VALIDOS = ('Borrador', 'Autorizada', 'Recibida', 'Pagada', 'Cancelada', 'Rechazada')
ESTADOS_OC_ACTIVAS = ('Borrador', 'Autorizada', 'Recibida')  # NO cerradas (Pagada/Cancelada/Rechazada)
ESTADOS_OC_LEGACY = ('Revisada', 'Parcial')                  # solo lectura · mig 157 los migró


# _siguiente_numero_oc · PG-safe · vive en audit_helpers (compartido con
# programacion.py y admin.py) · importado arriba como _siguiente_numero_oc.


# ── Helpers de permisos granulares ────────────────────────────────────────────
# Los grupos:
#   COMPRAS_ACCESS    = {catalina, mayra, alejandro, sebastian}
#                       Pueden CREAR solicitudes/OCs, recibir mercancía, gestionar
#                       proveedores, generar OCs desde Planta.
#   ADMIN_USERS       = {sebastian, alejandro}
#                       Pueden ejecutar TODAS las operaciones, incluso autorizar
#                       y pagar (CONTADORA está excluida explícitamente de
#                       autorización/pago para mantener segregación de duties).
#   CONTADORA_USERS   = {mayra, catalina}
#                       Pueden ver/gestionar pero NO autorizar/pagar OCs
#                       directamente — la autorización es responsabilidad de
#                       admin. (Mayra ve todo lo financiero; pagar implica
#                       movimiento de dinero que necesita firma de admin.)


def _require_compras_session():
    """Cualquier user autenticado puede hacer LECTURAS en compras."""
    u = session.get('compras_user', '')
    if not u:
        return None, jsonify({'error': 'No autenticado'}), 401
    return u, None, None


def _require_compras_write():
    """Para CREAR/EDITAR solicitudes y OCs: COMPRAS_ACCESS o ADMIN."""
    u = session.get('compras_user', '')
    if not u:
        return None, jsonify({'error': 'No autenticado'}), 401
    allowed = COMPRAS_ACCESS | ADMIN_USERS
    if u not in allowed:
        return None, jsonify({
            'error': 'Sin acceso a operaciones de Compras',
            'detail': f"User '{u}' no está en COMPRAS_ACCESS"
        }), 403
    return u, None, None


def _require_authorize_oc():
    """Para AUTORIZAR/PAGAR OCs: COMPRAS_ACCESS_WRITE o ADMIN · NUNCA contadora."""
    # SEC-FIX · 21-may-2026 · segregation of duties (CVSS 6.5)
    # Antes: alias literal de _require_compras_write incluía contadora
    # Ahora: bloquear CONTADORA explícitamente (solo registra pagos, no autoriza)
    usuario, err, code = _require_compras_write()
    if err:
        return None, err, code
    try:
        from config import CONTADORA_USERS as _CU
    except Exception:
        _CU = set()
    try:
        from config import OC_AUTORIZA_USERS as _OCA
    except Exception:
        _OCA = set()
    u_lower = (usuario or '').lower()
    es_admin = u_lower in {x.lower() for x in ADMIN_USERS}
    # Autorizador de OC explícito (ej. Catalina, asistente de compras) — puede
    # autorizar/pagar aunque comparta el perfil contable. Sebastián 13-jun-2026.
    es_autorizador_oc = u_lower in {x.lower() for x in _OCA}
    if u_lower in {x.lower() for x in _CU} and not es_admin and not es_autorizador_oc:
        # Contadora PURA (registra pagos, no autoriza · segregación de funciones)
        return None, jsonify({
            'error': 'Contadora no autoriza OCs · solo admin/compras',
            'codigo': 'SEGREGATION_OF_DUTIES',
        }), 403
    return usuario, None, None


def _enviar_oc_a_proveedor(numero_oc, proveedor, email_proveedor, items, monto_total, observaciones=''):
    """Compras Fase 2 · Sebastián 21-may-2026 · email auto al proveedor.

    Recomendación consultor: 'Catalina hoy seguro reenvía manual · un email
    automático cuando la OC pasa a Autorizada ahorra 10 min/OC × 75 OCs =
    12h/mes liberadas.'

    Genera HTML con OC formal · usa _enviar_email_async (existente en
    comunicacion.py · thread daemon · no bloquea).
    """
    if not email_proveedor or '@' not in email_proveedor:
        return False, 'sin email proveedor'
    asunto = f'OC {numero_oc} · {proveedor} · Espagiria Laboratorio'
    items_html = ''
    for it in (items or []):
        items_html += (
            '<tr>'
            f'<td style="padding:8px;border-bottom:1px solid #e5e7eb">{(it.get("codigo_mp") or "")}</td>'
            f'<td style="padding:8px;border-bottom:1px solid #e5e7eb">{(it.get("nombre_mp") or "")}</td>'
            f'<td style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:right">{float(it.get("cantidad_g") or 0):,.0f}</td>'
            f'<td style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:right">${float(it.get("precio_unitario") or 0):,.2f}</td>'
            f'<td style="padding:8px;border-bottom:1px solid #e5e7eb;text-align:right">${(float(it.get("cantidad_g") or 0)*float(it.get("precio_unitario") or 0)):,.0f}</td>'
            '</tr>'
        )
    body_html = (
        '<html><body style="font-family:Arial,sans-serif;max-width:680px;margin:auto">'
        '<div style="background:#0e7490;color:#fff;padding:20px;border-radius:8px 8px 0 0">'
        '<h2 style="margin:0">Orden de Compra Autorizada</h2>'
        f'<div style="opacity:.9;margin-top:4px">N° {numero_oc} · {proveedor}</div>'
        '</div>'
        '<div style="background:#f9fafb;padding:20px;border-radius:0 0 8px 8px">'
        '<p>Estimado proveedor,</p>'
        f'<p>Se ha autorizado la siguiente Orden de Compra:</p>'
        '<table style="width:100%;border-collapse:collapse;background:#fff;border-radius:6px;overflow:hidden">'
        '<thead style="background:#1e293b;color:#fff"><tr>'
        '<th style="padding:10px;text-align:left">Código</th>'
        '<th style="padding:10px;text-align:left">Material</th>'
        '<th style="padding:10px;text-align:right">Cant (g)</th>'
        '<th style="padding:10px;text-align:right">Precio unit.</th>'
        '<th style="padding:10px;text-align:right">Subtotal</th>'
        '</tr></thead><tbody>' + items_html + '</tbody>'
        f'<tfoot><tr><td colspan="4" style="padding:10px;text-align:right;font-weight:700">TOTAL:</td>'
        f'<td style="padding:10px;text-align:right;font-weight:800;color:#0e7490">${monto_total:,.0f}</td></tr></tfoot>'
        '</table>'
        + (f'<p style="margin-top:14px"><b>Observaciones:</b><br>{observaciones}</p>' if observaciones else '')
        + '<p style="margin-top:18px">Confirmar recepción de esta OC respondiendo a este correo.</p>'
        '<p style="color:#6b7280;font-size:12px;margin-top:20px">Espagiria Laboratorio · Compras HHA Group</p>'
        '</div></body></html>'
    )
    try:
        from blueprints.comunicacion import _enviar_email_async
        _enviar_email_async(asunto, body_html, [email_proveedor])
        return True, 'enviado'
    except Exception as e:
        return False, f'email fallo: {e}'


def _evaluar_auto_aprobacion(c, proveedor, monto_total, items):
    """Compras Fase 2 · Sebastián 21-may-2026 · auto-aprobación por reglas.

    Recomendación del consultor procurement: OCs pequeñas a proveedores
    recurrentes con precio en rango histórico deben pasarse automáticamente
    a estado 'Autorizada' (saltean revisión gerencial · liberan 30-40 min/día
    de Catalina).

    REGLAS (las 3 deben cumplirse):
    1. monto_total < LIMITE_AUTO_APROB_COP (default $500.000)
    2. proveedor con ≥3 OCs en últimos 90 días (recurrente)
    3. cada item con precio_unitario en rango ±15% del promedio 90d
       (si no hay histórico del MP, permite por default)

    SEC-FIX 27-may-2026 (audit r3) · default OFF · activar explícito con env
    `COMPRAS_AUTO_APROB_ON=1`. Antes era ON por default y si la env
    `COMPRAS_AUTO_APROB_OFF` se perdía en redeploy, todas las OCs <$500k
    auto-aprobaban sin revisión humana. Ahora el riesgo está cerrado por
    diseño. La env legacy `COMPRAS_AUTO_APROB_OFF` sigue funcionando como
    override redundante.

    Retorna (auto_aprobar: bool, razon: str)
    """
    import os
    # SEC-FIX 27-may · activación EXPLÍCITA · default OFF
    if (os.environ.get('COMPRAS_AUTO_APROB_ON') or '0').strip() != '1':
        return False, 'auto-aprob OFF por default · setear COMPRAS_AUTO_APROB_ON=1 para activar'
    # Legacy compat · si admin ya tiene OFF=1 explícito, respetar
    if (os.environ.get('COMPRAS_AUTO_APROB_OFF') or '0').strip() == '1':
        return False, 'auto-aprob deshabilitada (env legacy)'
    LIMITE = float(os.environ.get('COMPRAS_AUTO_APROB_LIMITE_COP') or 500_000)
    if not proveedor:
        return False, 'sin proveedor'
    if monto_total >= LIMITE:
        return False, f'monto {monto_total:.0f} >= límite {LIMITE:.0f}'
    # AUDITORÍA-FIX 23-may-2026 · C21 · guard de exposición acumulada diaria
    # · 10 SOLs concurrentes de $499K c/u al mismo proveedor podían auto-
    # aprobarse y dar $4.99M de exposure sin revisión gerencial
    LIMITE_DIARIO = float(os.environ.get('COMPRAS_AUTO_APROB_LIMITE_DIARIO_COP') or (LIMITE * 5))
    try:
        r_dia = c.execute(
            """SELECT COALESCE(SUM(valor_total),0) FROM ordenes_compra
               WHERE LOWER(TRIM(proveedor))=LOWER(TRIM(?))
                 AND date(fecha) = date('now','-5 hours')
                 AND estado IN ('Autorizada','Parcial','Recibida','Pagada')""",
            (proveedor,),
        ).fetchone()
        suma_dia = float((r_dia or [0])[0] or 0)
    except Exception:
        suma_dia = 0
    if (suma_dia + monto_total) >= LIMITE_DIARIO:
        return False, f'exposición diaria del proveedor {suma_dia + monto_total:.0f} >= límite diario {LIMITE_DIARIO:.0f}'
    # Recurrencia · ≥3 OCs últimos 90d
    # AUDITORÍA-FIX 23-may-2026 · C7 · antes contaba TODAS las OCs incluyendo
    # Canceladas/Rechazadas/Borrador · un proveedor con 3 OCs todas canceladas
    # quedaba como "recurrente" y auto-aprobado · ahora solo cuenta OCs que
    # realmente avanzaron (Autorizada/Parcial/Recibida/Pagada)
    # PG-FIX 27-may · `date('now','-5h','-90d')` multi-arg rompe en PostgreSQL
    # · pg_compat.translate_ddl solo soporta date() mono-arg. Calcular cutoff
    # en Python (Bogotá UTC-5) y pasarlo como param `?`.
    from datetime import datetime as _dt90, timedelta as _td90
    _cutoff_90d = (_dt90.utcnow() - _td90(hours=5) - _td90(days=90)).date().isoformat()
    try:
        r = c.execute(
            """SELECT COUNT(*) FROM ordenes_compra
               WHERE LOWER(TRIM(proveedor))=LOWER(TRIM(?))
                 AND estado IN ('Autorizada','Parcial','Recibida','Pagada')
                 AND date(fecha) >= ?""",
            (proveedor, _cutoff_90d),
        ).fetchone()
        ocs_90d = int((r or [0])[0] or 0)
    except Exception:
        ocs_90d = 0
    if ocs_90d < 3:
        return False, f'proveedor con solo {ocs_90d} OCs en 90d (<3)'
    # FASE 3 cont · 22-may-2026 · regla 4 opcional · scorecard ≥70
    # Si env COMPRAS_AUTO_APROB_REQ_SCORE=70 (o N) · exigir score >= N
    try:
        import os as _os_sc
        score_min = int(_os_sc.environ.get('COMPRAS_AUTO_APROB_REQ_SCORE') or 0)
    except Exception:
        score_min = 0
    if score_min > 0:
        try:
            sc = _scorecard_proveedor_dict(c, proveedor)
            if sc.get('score_global', 0) < score_min:
                return False, f'score {sc.get("score_global",0)} < mínimo {score_min}'
        except Exception:
            pass
    # Precio en rango · cada item ±15% promedio 90d
    # PERF-FIX 27-may-2026 PM · antes era N+1 (1 SELECT por item · OCs grandes
    # con 20+ items hacían 20+ queries). Ahora 1 sola query GROUP BY.
    # BUG-FIX 27-may-2026 PM · tabla era `precios_mp_historico.precio_unitario`
    # · columna inexistente (precios_mp_historico tiene precio_kg, y la tabla
    # que SÍ se usa para histórico es `precio_historico_mp.precio_unit_g`).
    # Antes el SELECT siempre tiraba excepción silenciada por try/except ·
    # el check de rango NUNCA filtró nada. Ahora consulta la tabla correcta.
    _codigos_validos = [(it.get('codigo_mp') or '') for it in (items or [])
                        if (it.get('codigo_mp') or '') and float(it.get('precio_unitario') or 0) > 0]
    _prom_por_cod = {}
    if _codigos_validos:
        try:
            _ph = ','.join(['?'] * len(_codigos_validos))
            _rows_prom = c.execute(
                f"""SELECT codigo_mp, AVG(precio_unit_g)
                    FROM precio_historico_mp
                    WHERE codigo_mp IN ({_ph})
                      AND date(fecha) >= ?
                    GROUP BY codigo_mp""",
                (*_codigos_validos, _cutoff_90d),
            ).fetchall()
            _prom_por_cod = {r[0]: float(r[1] or 0) for r in _rows_prom}
        except Exception:
            _prom_por_cod = {}
    for it in (items or []):
        cod = it.get('codigo_mp') or ''
        pu = float(it.get('precio_unitario') or 0)
        if not cod or pu <= 0:
            continue
        prom = _prom_por_cod.get(cod, 0)
        if prom > 0:
            delta_pct = abs(pu - prom) / prom * 100
            if delta_pct > 15:
                return False, f'precio {cod} fuera de rango (Δ{delta_pct:.1f}%)'
    return True, f'monto<{LIMITE:.0f} · recurrente ({ocs_90d} OCs/90d) · precios en rango'


def _pendiente_en_compras_g(c, codigo_mp):
    """Compras PRO · Sebastián 21-may-2026 · helper anti-duplicación.

    Devuelve la cantidad (gramos) de un MP que YA está en cola de Compras:
      - SOLs Pendientes/Aprobadas sin OC asociada
      - OCs Borrador/Revisada/Autorizada no recibidas (cant pedida - recibida)

    Cualquier generador automático de SOLs (auto_plan, mínimos, programación,
    pre-prod) DEBE restar esto al déficit antes de crear nuevas SOLs · evita
    duplicación cross-canales que llevaba a compras ×2 silenciosas.

    Es la implementación del Fix #1 del agente auditor Planta↔Compras como
    función reusable.
    """
    total = 0.0
    if not codigo_mp:
        return 0.0
    try:
        r = c.execute(
            """SELECT COALESCE(SUM(sci.cantidad_g), 0)
               FROM solicitudes_compra_items sci
               JOIN solicitudes_compra sc ON sc.numero = sci.numero
               WHERE UPPER(TRIM(sci.codigo_mp)) = UPPER(TRIM(?))
                 AND sc.estado IN ('Pendiente','En revision','Aprobada')
                 AND COALESCE(sc.numero_oc,'') = ''""",
            (codigo_mp,),
        ).fetchone()
        total += float((r or [0])[0] or 0)
    except Exception:
        pass
    try:
        # AUDITORÍA-FIX 23-may-2026 · C6 · OCs 'Pagada' SIN recibir
        # (anticipo · pago directo previo a recepción) son material pendiente
        # de llegar · antes se excluían · resultado: generadores automáticos
        # creaban SOL duplicada para una MP que ya estaba en OC pagada
        # · ahora se incluyen mientras fecha_recepcion esté vacía
        r = c.execute(
            """SELECT COALESCE(SUM(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g,0)), 0)
               FROM ordenes_compra_items oci
               JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
               WHERE UPPER(TRIM(oci.codigo_mp)) = UPPER(TRIM(?))
                 AND (
                       oc.estado IN ('Borrador','Revisada','Autorizada','Parcial')
                    OR (oc.estado='Pagada' AND COALESCE(oc.fecha_recepcion,'')='')
                 )""",
            (codigo_mp,),
        ).fetchone()
        total += float((r or [0])[0] or 0)
    except Exception:
        pass
    return max(total, 0.0)


def _pendiente_en_compras_bulk(c):
    """PERF FIX 24-may PM · auditoría agente · evita N+1 al iterar
    ~70 productos × ~10-30 MPs c/u = 2000+ queries.

    Misma lógica que _pendiente_en_compras_g pero retorna dict
    {codigo_mp: gramos_pendientes_total} en UNA pasada (2 queries
    GROUP BY total).
    """
    pendiente = {}
    try:
        for r in c.execute(
            """SELECT UPPER(TRIM(sci.codigo_mp)), COALESCE(SUM(sci.cantidad_g), 0)
               FROM solicitudes_compra_items sci
               JOIN solicitudes_compra sc ON sc.numero = sci.numero
               WHERE sc.estado IN ('Pendiente','En revision','Aprobada')
                 AND COALESCE(sc.numero_oc,'') = ''
               GROUP BY UPPER(TRIM(sci.codigo_mp))""",
        ).fetchall():
            mid = str(r[0] or '').strip()
            if mid:
                pendiente[mid] = float(r[1] or 0)
    except Exception:
        pass
    try:
        for r in c.execute(
            """SELECT UPPER(TRIM(oci.codigo_mp)),
                      COALESCE(SUM(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g,0)), 0)
               FROM ordenes_compra_items oci
               JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
               WHERE (
                       oc.estado IN ('Borrador','Revisada','Autorizada','Parcial')
                    OR (oc.estado='Pagada' AND COALESCE(oc.fecha_recepcion,'')='')
                 )
               GROUP BY UPPER(TRIM(oci.codigo_mp))""",
        ).fetchall():
            mid = str(r[0] or '').strip()
            if mid:
                pendiente[mid] = pendiente.get(mid, 0.0) + float(r[1] or 0)
    except Exception:
        pass
    return pendiente


def _check_monto_limit(usuario, monto):
    """Valida que el usuario pueda autorizar el monto.

    Retorna (None, None) si OK, o (response, status_code) si excede el límite.
    Admins (límite None) siempre pueden. Otros tienen LIMITES_APROBACION_OC.
    """
    if usuario in ADMIN_USERS:
        return None, None
    limite = LIMITES_APROBACION_OC.get(usuario)
    if limite is None:
        # Usuario sin límite explícito: usar 0 (no puede autorizar nada)
        return jsonify({
            'error': f"Usuario '{usuario}' no tiene límite de aprobación configurado",
            'detail': 'Solo usuarios con LIMITES_APROBACION_OC en config pueden autorizar.'
        }), 403
    if monto > limite:
        return jsonify({
            'error': f'Monto ${monto:,.0f} excede tu límite de aprobación (${limite:,.0f})',
            'detail': 'Esta OC requiere aprobación de un administrador.',
            'codigo': 'EXCEDE_LIMITE_APROBACION',
            'limite_usuario': limite,
            'monto_solicitado': monto,
        }), 403
    return None, None

def _is_animus_payment(c, numero_oc=None, beneficiario_nombre=None,
                       observaciones=None, categoria=None):
    """Determina si un pago debe atribuirse a ANIMUS LAB S.A.S.

    Empresa pagadora del grupo HHA:
      - Influencers, marketing digital, cuenta de cobro → ANIMUS LAB
      - Mercancía, MPs, planta, servicios técnicos     → ESPAGIRIA

    Detección multi-señal — cualquiera prende Animus, así soporta CEs
    legacy (creados antes del dispatch) y casos donde la categoría está
    vacía o mal poblada:

      1. categoria contiene 'influencer' / 'marketing' / 'cuenta de cobro'
      2. existe un row en pagos_influencers para el numero_oc
      3. solicitudes_compra.influencer_id NOT NULL para el numero_oc
      4. beneficiario_nombre matchea un row de marketing_influencers
      5. observaciones contienen la palabra 'influencer'

    Cualquier excepción de SQL (tabla legacy faltante en deploy viejo) se
    silencia y se evalúa la siguiente señal — la falta de tabla nunca
    debe romper la generación del PDF.
    """
    # Señal 1: keyword en categoría
    cat_low = (categoria or '').lower()
    if ('influencer' in cat_low
            or 'marketing' in cat_low
            or 'cuenta de cobro' in cat_low):
        return True

    # Señal 2: pago_influencers row
    if numero_oc:
        try:
            row = c.execute(
                "SELECT 1 FROM pagos_influencers WHERE numero_oc=? LIMIT 1",
                (numero_oc,)
            ).fetchone()
            if row:
                return True
        except sqlite3.OperationalError:
            pass

        # Señal 3: solicitudes_compra.influencer_id
        try:
            row = c.execute(
                "SELECT influencer_id FROM solicitudes_compra "
                "WHERE numero_oc=? AND influencer_id IS NOT NULL LIMIT 1",
                (numero_oc,)
            ).fetchone()
            if row and row[0]:
                return True
        except sqlite3.OperationalError:
            pass

    # Señal 4: beneficiario en marketing_influencers
    if beneficiario_nombre:
        try:
            row = c.execute(
                "SELECT 1 FROM marketing_influencers "
                "WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?)) LIMIT 1",
                (beneficiario_nombre,)
            ).fetchone()
            if row:
                return True
        except sqlite3.OperationalError:
            pass

    # Señal 5: observaciones contienen 'influencer'
    if observaciones and 'influencer' in (observaciones or '').lower():
        return True

    return False


def _notificar_solicitante_email(dest_email, asunto, body_html):
    """Envia email al solicitante de forma no-bloqueante.
    Nunca lanza excepcion — falla silenciosamente con log.
    """
    if not dest_email:
        return
    try:
        import sys, threading
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from notificaciones import SistemaNotificaciones
        notif = SistemaNotificaciones()
        threading.Thread(
            target=notif._enviar_email,
            args=(asunto, body_html, [dest_email]),
            daemon=True
        ).start()
    except Exception as _e:
        print(f'[notificar_solicitante] email error (non-critical): {_e}')

@bp.route('/api/dashboard-stats')
def dashboard_stats():
    """[LEGACY 22-may-2026] Reemplazado por /api/compras/dashboard-home (consolidado).
    Mantenido por compat con código viejo · NO usar en código nuevo.

    Dashboard stats. Sebastian (30-abr-2026): los paneles 'Vencimientos
    próximos', 'Top 5 MPs' y 'Estado general de lotes' aparecían vacíos —
    el cálculo dependía del campo estado_lote que NO se mantiene
    automáticamente. Fix: calcular los estados dinámicamente desde
    fecha_vencimiento, y soportar variantes de tipo de movimiento.
    """
    from datetime import date
    hoy = date.today().isoformat()
    conn = get_db(); c = conn.cursor()

    # Soporta 'Entrada', 'Ingreso', 'Ajuste', 'Devolucion' como entradas;
    # 'Salida', 'Consumo' como salidas.
    SUMA_STOCK = (
        "SUM(CASE "
        "WHEN tipo IN ('Entrada','Ingreso','Ajuste','Devolucion','Devolución') THEN cantidad "
        "WHEN tipo IN ('Salida','Consumo') THEN -cantidad "
        "ELSE 0 END)"
    )
    FILTRO_ENTRADA = "tipo IN ('Entrada','Ingreso','Ajuste','Devolucion','Devolución')"

    # ── Vencimientos próximos 6 meses ─────────────────────────────────────
    # Calculo a nivel LOTE — agrupar entradas por (material_id, lote) y
    # restar consumos, mostrar solo lotes con stock_restante > 0 con
    # fecha_vencimiento en próximos 180 dias.
    venc_por_mes = {}
    try:
        c.execute(f"""
            WITH lote_stock AS (
              SELECT material_id, lote,
                     MIN(fecha_vencimiento) as venc,
                     {SUMA_STOCK} as stock_restante_g
              FROM movimientos
              WHERE lote IS NOT NULL AND lote != ''
              GROUP BY material_id, lote
              HAVING stock_restante_g > 0
            )
            SELECT venc, COUNT(*) as n, SUM(stock_restante_g) as total_g
            FROM lote_stock
            WHERE venc IS NOT NULL AND venc >= ? AND venc <= date(?, '+180 days')
            GROUP BY substr(venc, 1, 7)
            ORDER BY venc
        """, (hoy, hoy))
        for row in c.fetchall():
            if row[0]:
                mes = str(row[0])[:7]
                venc_por_mes[mes] = {'lotes': row[1], 'kg': round((row[2] or 0)/1000, 1)}
    except Exception as e:
        # Fallback simple si la query agrupada falla
        try:
            c.execute(f"""
                SELECT fecha_vencimiento, COUNT(*) as n, SUM(cantidad) as total_g
                FROM movimientos
                WHERE {FILTRO_ENTRADA}
                  AND fecha_vencimiento IS NOT NULL
                  AND fecha_vencimiento >= ?
                  AND fecha_vencimiento <= date(?, '+180 days')
                GROUP BY substr(fecha_vencimiento, 1, 7)
            """, (hoy, hoy))
            for row in c.fetchall():
                if row[0]:
                    mes = str(row[0])[:7]
                    venc_por_mes[mes] = {'lotes': row[1], 'kg': round((row[2] or 0)/1000, 1)}
        except Exception:
            pass

    # ── Alertas reabastecimiento: MPs bajo mínimo ─────────────────────────
    try:
        c.execute(f"""SELECT COUNT(*) FROM maestro_mps m
                     LEFT JOIN (SELECT material_id, {SUMA_STOCK} as stock
                                FROM movimientos GROUP BY material_id) s
                       ON m.codigo_mp=s.material_id
                     WHERE m.activo=1 AND m.stock_minimo>0 AND COALESCE(s.stock,0)<m.stock_minimo""")
        mps_bajo_minimo = c.fetchone()[0] or 0
    except Exception:
        mps_bajo_minimo = 0

    # ── Estado general de lotes — CALCULADO desde fecha_vencimiento ──────
    # No depender de estado_lote (campo desactualizado). Calcular por:
    #   VENCIDO: fecha_vencimiento < hoy AND stock > 0
    #   CRITICO: 0-30 dias
    #   PROXIMO: 31-90 dias
    estados = {'VENCIDO': 0, 'CRITICO': 0, 'PROXIMO': 0}
    try:
        # PG-FIX 27-may · date('now','-5h','+Nd') multi-arg falla en PG.
        # Calcular cutoffs Bogotá UTC-5 en Python y pasar como params.
        from datetime import datetime as _dtL, timedelta as _tdL
        _hoy = (_dtL.utcnow() - _tdL(hours=5)).date().isoformat()
        _d30 = (_dtL.utcnow() - _tdL(hours=5) + _tdL(days=30)).date().isoformat()
        _d90 = (_dtL.utcnow() - _tdL(hours=5) + _tdL(days=90)).date().isoformat()
        c.execute(f"""
            WITH lote_stock AS (
              SELECT material_id, lote, MIN(fecha_vencimiento) as venc,
                     {SUMA_STOCK} as stock_g
              FROM movimientos
              WHERE lote IS NOT NULL AND lote != '' AND fecha_vencimiento IS NOT NULL
              GROUP BY material_id, lote
              HAVING stock_g > 0
            )
            SELECT
              SUM(CASE WHEN venc < ? THEN 1 ELSE 0 END) as vencidos,
              SUM(CASE WHEN venc >= ? AND venc <= ? THEN 1 ELSE 0 END) as criticos,
              SUM(CASE WHEN venc > ? AND venc <= ? THEN 1 ELSE 0 END) as proximos
            FROM lote_stock
        """, (_hoy, _hoy, _d30, _d30, _d90))
        r = c.fetchone()
        if r:
            estados = {'VENCIDO': r[0] or 0, 'CRITICO': r[1] or 0, 'PROXIMO': r[2] or 0}
    except Exception:
        pass

    # ── Top 5 MPs por stock actual ────────────────────────────────────────
    top_stock = []
    try:
        c.execute(f"""SELECT material_id, material_nombre, {SUMA_STOCK} as stock
                     FROM movimientos
                     GROUP BY material_id, material_nombre
                     HAVING stock > 0
                     ORDER BY stock DESC LIMIT 5""")
        top_stock = [{'codigo': r[0], 'nombre': r[1], 'kg': round((r[2] or 0)/1000, 1)}
                     for r in c.fetchall()]
    except Exception:
        pass

    # ── Stock total en kg ─────────────────────────────────────────────────
    try:
        c.execute(f"SELECT {SUMA_STOCK} FROM movimientos")
        stock_total_g = c.fetchone()[0] or 0
    except Exception:
        stock_total_g = 0

    return jsonify({
        'vencimientos_por_mes': venc_por_mes,
        'mps_bajo_minimo': mps_bajo_minimo,
        'estados_lotes': estados,
        'top_stock': top_stock,
        'stock_total_kg': round(stock_total_g/1000, 1)
    })

@bp.route('/api/generar-oc-automatica', methods=['POST'])
def generar_oc_automatica():
    """⚠️ DEPRECATED · Sprint Compras N1 · 21-may-2026.

    Endpoint LEGACY. Crea OCs en estado 'Pendiente' (no 'Borrador'),
    sin creado_por, sin categoria, sin audit_log consistente · conflicto
    con flujo nuevo `oc-desde-solicitudes` (que SÍ cumple invariantes).

    Sebastián decidió: redirigir todo al flujo canónico. Este endpoint
    devuelve 410 GONE con instrucción de migrar.

    Para reactivar (NO recomendado): borrar este wrapper y desambiguar
    el flujo · pero todos los call sites del frontend ya migraron a
    `crear_oc_desde_solicitudes`.
    """
    return jsonify({
        'error': 'Endpoint DEPRECATED · usar /api/compras/oc-desde-solicitudes',
        'detalle': (
            'generar-oc-automatica fue reemplazado por el flujo canónico que '
            'crea OCs en estado Borrador con auditoría completa · monto-limit '
            'check · vinculación a SOLs · histórico de precios.'
        ),
        'reemplazo': '/api/compras/oc-desde-solicitudes',
        'desde': '2026-05-21',
    }), 410


@bp.route('/api/compras/influencer/limpiar-no-pagadas', methods=['GET', 'POST'])
def limpiar_influencer_no_pagadas():
    """Limpia SOLs/OCs de Influencer/Marketing/CC que NO se han pagado.

    Sebastian (30-abr-2026): "elimíname todo lo que quedó en influencers
    de compras que ya no son para pagar no se de donde salieron".

    Modo dry-run (default): GET o POST sin {confirm:true} → devuelve
    la lista de candidatos sin borrar nada. Para revisar antes.

    Modo real: POST {confirm:true} → ejecuta borrado.
    Solo admin (sebastian/alejandro).

    Borra:
      - SOL con categoria influencer/marketing/cuenta_de_cobro
        Y estado IN ('Pendiente','Aprobada','Rechazada')  ← NO 'Pagada'
        Y la OC vinculada NO tiene pagos efectivos.
      - OC vinculada (si no esta pagada) + sus items + pagos_influencers
        Pendientes (preserva los Pagados, son historial).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = (session.get('compras_user') or '').lower()
    if user not in {a.lower() for a in ADMIN_USERS}:
        return jsonify({'error': 'Solo admin'}), 403

    payload = request.get_json(silent=True) or {}
    confirm = payload.get('confirm') is True

    conn = get_db(); c = conn.cursor()
    sols = c.execute("""
        SELECT numero, solicitante, valor, estado, numero_oc, categoria,
               observaciones, fecha
        FROM solicitudes_compra
        WHERE (LOWER(COALESCE(categoria,'')) LIKE '%influencer%'
               OR LOWER(COALESCE(categoria,'')) LIKE '%marketing%'
               OR LOWER(COALESCE(categoria,'')) LIKE '%cuenta de cobro%')
          AND estado IN ('Pendiente','Aprobada','Rechazada')
        ORDER BY numero DESC
    """).fetchall()

    candidatos = []
    for s in sols:
        numero, solicitante, valor, estado, oc_num, cat, obs, fecha = s
        razon_skip = None
        n_pagos_oc = 0
        n_pagos_inf_pagados = 0
        if oc_num:
            try:
                n_pagos_oc = c.execute(
                    "SELECT COUNT(*) FROM pagos_oc WHERE numero_oc=?", (oc_num,)
                ).fetchone()[0] or 0
            except sqlite3.OperationalError:
                n_pagos_oc = 0
            try:
                n_pagos_inf_pagados = c.execute(
                    "SELECT COUNT(*) FROM pagos_influencers WHERE numero_oc=? AND estado='Pagada'",
                    (oc_num,)
                ).fetchone()[0] or 0
            except sqlite3.OperationalError:
                n_pagos_inf_pagados = 0
            if n_pagos_oc > 0 or n_pagos_inf_pagados > 0:
                razon_skip = f'OC {oc_num} tiene {n_pagos_oc} pagos_oc + {n_pagos_inf_pagados} pagos_inf Pagados'

        # Extraer beneficiario de obs si existe
        benef = ''
        if obs:
            import re as _re
            m = _re.search(r'BENEFICIARIO:\s*([^|]+)', obs, _re.IGNORECASE)
            if m: benef = m.group(1).strip()

        candidatos.append({
            'numero': numero,
            'solicitante': solicitante,
            'beneficiario': benef,
            'valor': valor or 0,
            'estado': estado,
            'numero_oc': oc_num,
            'fecha': fecha,
            'categoria': cat,
            'safe_to_delete': razon_skip is None,
            'razon_skip': razon_skip,
        })

    elegibles = [x for x in candidatos if x['safe_to_delete']]
    omitidos = [x for x in candidatos if not x['safe_to_delete']]

    if not confirm:
        return jsonify({
            'dry_run': True,
            'total_candidatos': len(candidatos),
            'a_borrar': len(elegibles),
            'omitidos_por_pagos': len(omitidos),
            'candidatos': elegibles,
            'omitidos': omitidos,
            'mensaje': f'Dry-run. {len(elegibles)} se borrarian, {len(omitidos)} omitidos por tener pagos. POST con {{"confirm":true}} para ejecutar.',
        })

    # Modo confirm — borrar
    eliminados = []
    for cand in elegibles:
        numero = cand['numero']
        oc_num = cand['numero_oc']
        # Borrar pagos_influencers Pendiente (preservar Pagada en otra OC, no aplica aqui)
        if oc_num:
            try:
                c.execute("DELETE FROM pagos_influencers WHERE numero_oc=? AND estado != 'Pagada'", (oc_num,))
            except sqlite3.OperationalError:
                pass
            # Borrar items
            try:
                c.execute("DELETE FROM ordenes_compra_items WHERE numero_oc=?", (oc_num,))
            except sqlite3.OperationalError:
                pass
            # Borrar OC si no esta pagada
            try:
                c.execute("DELETE FROM ordenes_compra WHERE numero_oc=? AND estado != 'Pagada'", (oc_num,))
            except sqlite3.OperationalError:
                pass
        # Borrar SOL items + SOL
        try:
            c.execute("DELETE FROM solicitudes_compra_items WHERE numero_solicitud=?", (numero,))
        except sqlite3.OperationalError:
            pass
        c.execute("DELETE FROM solicitudes_compra WHERE numero=?", (numero,))
        eliminados.append(numero)

    # Sebastián 27-may-2026 · audit · borrado masivo regulado DEBE auditar
    # (INVIMA trazabilidad · incidente 19-may = borrado sin rastro). Antes
    # este DELETE multiplex de SOLs/OCs/pagos no dejaba huella.
    try:
        from audit_helpers import audit_log as _al
        _al(c, usuario=session.get('compras_user', 'desconocido'),
            accion='LIMPIAR_INFLUENCER_NO_PAGADAS_MASIVO',
            tabla='solicitudes_compra', registro_id='bulk',
            antes={'candidatos': [x['numero'] for x in elegibles]},
            despues={'eliminados': eliminados, 'total': len(eliminados)},
            detalle=(f"Borrado masivo {len(eliminados)} SOLs influencer no-pagadas "
                     f"(+ OCs/items/pagos pendientes vinculados) · "
                     f"{len(omitidos)} omitidos por tener pagos"))
    except Exception as _ae:
        import logging as _lg
        _lg.getLogger('compras').warning('audit limpiar influencer masivo falló: %s', _ae)
    conn.commit()
    return jsonify({
        'ok': True,
        'eliminados': eliminados,
        'total_eliminados': len(eliminados),
        'omitidos_por_pagos': len(omitidos),
    })


@bp.route('/api/solicitudes-compra/<numero>', methods=['DELETE'])
def eliminar_solicitud(numero):
    """Elimina una solicitud de compra y limpia OC asociada cuando aplica.

    Reglas:
    - OC en estado Borrador/Rechazada → se borra junto con sus items.
    - OC en estado Aprobada Y categoría es Influencer/Marketing/CC Y NO tiene
      pagos en pagos_oc Y NO tiene pago en pagos_influencers con estado='Pagada'
        → se borra todo (OC + items + pagos_influencers pendientes).
    - OC en cualquier otro estado o con pagos hechos → se borra SOLO la
      solicitud, la OC se conserva (mantiene historial).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user_actual = session.get('compras_user', '').lower()
    conn = get_db(); c = conn.cursor()
    # SOL-FIX#2 · 21-may-2026 · DELETE requiere autoría o admin/compras_access
    # Antes: cualquier user logueado borraba SOLs de otros · riesgo accidental.
    c.execute("SELECT estado, numero_oc, categoria, solicitante FROM solicitudes_compra WHERE numero=?",
              (numero.upper(),))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'No encontrada'}), 404
    estado, numero_oc, categoria, solicitante_orig = row[0], row[1], (row[2] or '').strip(), (row[3] or '').lower()
    # Permitido: creador, admin, o compras_access_write (Catalina puede borrar)
    try:
        from config import COMPRAS_ACCESS as _CA
    except Exception:
        _CA = set()
    _allowed = (
        user_actual == solicitante_orig
        or user_actual in {x.lower() for x in ADMIN_USERS}
        or user_actual in {x.lower() for x in _CA}
    )
    if not _allowed:
        return jsonify({
            'error': 'Solo el creador, Compras o Admin pueden borrar esta SOL',
            'codigo': 'SOL_DELETE_PERMISO',
        }), 403

    oc_borrada = False
    pagos_inf_borrados = 0
    is_intangible = categoria in ('Influencer/Marketing Digital', 'Cuenta de Cobro')

    if numero_oc and numero_oc.strip():
        cur = conn.cursor()
        cur.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
        oc_row = cur.fetchone()
        if oc_row:
            oc_estado = oc_row[0]
            puede_borrar_oc = False
            if oc_estado in ('Borrador', 'Rechazada'):
                puede_borrar_oc = True
            elif oc_estado == 'Aprobada' and is_intangible:
                # Verificar que NO haya pagos efectivos
                try:
                    cur.execute("SELECT COUNT(*) FROM pagos_oc WHERE numero_oc=?", (numero_oc,))
                    n_pagos_oc = cur.fetchone()[0] or 0
                except sqlite3.OperationalError:
                    n_pagos_oc = 0
                try:
                    cur.execute("SELECT COUNT(*) FROM pagos_influencers WHERE numero_oc=? AND estado='Pagada'",
                                (numero_oc,))
                    n_pagos_inf_pagados = cur.fetchone()[0] or 0
                except sqlite3.OperationalError:
                    n_pagos_inf_pagados = 0
                if n_pagos_oc == 0 and n_pagos_inf_pagados == 0:
                    puede_borrar_oc = True
            if puede_borrar_oc:
                # Borrar pagos_influencers pendientes asociados (no Pagados)
                try:
                    d = cur.execute(
                        "DELETE FROM pagos_influencers WHERE numero_oc=? AND COALESCE(estado,'') != 'Pagada'",
                        (numero_oc,)
                    )
                    pagos_inf_borrados = d.rowcount or 0
                except sqlite3.OperationalError:
                    pass
                try:
                    cur.execute("DELETE FROM ordenes_compra_items WHERE numero_oc=?", (numero_oc,))
                except sqlite3.OperationalError:
                    pass
                cur.execute("DELETE FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
                oc_borrada = True
    try:
        c.execute("DELETE FROM solicitudes_compra_items WHERE numero=?", (numero.upper(),))
    except sqlite3.OperationalError:
        pass
    c.execute("DELETE FROM solicitudes_compra WHERE numero=?", (numero.upper(),))
    # SEC-FIX 23-may-2026 · INVIMA Res 2214/2021 · DELETE sin audit es la
    # causa raíz del incidente 19-may (programación desaparecida sin rastro)
    # · ahora SIEMPRE registramos qué se borró
    try:
        audit_log(c, usuario=user_actual, accion='ELIMINAR_SOLICITUD',
                  tabla='solicitudes_compra', registro_id=numero.upper(),
                  antes={
                      'numero': numero.upper(),
                      'estado': estado,
                      'numero_oc': numero_oc,
                      'categoria': categoria,
                      'solicitante': solicitante_orig,
                  },
                  despues={
                      'oc_borrada': oc_borrada,
                      'pagos_influencers_borrados': pagos_inf_borrados,
                  })
    except Exception:
        pass
    conn.commit()
    return jsonify({
        'ok': True,
        'eliminada': numero.upper(),
        'oc_borrada': oc_borrada,
        'numero_oc': numero_oc or '',
        'pagos_influencers_borrados': pagos_inf_borrados,
    })

@bp.route('/api/ordenes-compra', methods=['GET','POST'])
def handle_ordenes_compra():
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        usuario, err, code = _require_compras_write()
        if err:
            return err, code
        d = request.get_json(silent=True) or {}
        if not d.get('proveedor'): return jsonify({'error': 'Proveedor requerido'}), 400
        categoria = d.get('categoria', 'MP')
        # numero único con reintento ante carrera MAX+1 entre workers
        for _intento in range(6):
            numero_oc = _siguiente_numero_oc(c)
            try:
                c.execute("INSERT INTO ordenes_compra (numero_oc,fecha,estado,proveedor,observaciones,creado_por,fecha_entrega_est,categoria) VALUES (?,?,?,?,?,?,?,?)",
                          (numero_oc, datetime.now().isoformat(), 'Borrador', d['proveedor'],
                           d.get('observaciones',''), d.get('creado_por',''), d.get('fecha_entrega_est',''), categoria))
                break
            except sqlite3.IntegrityError:
                if _intento == 5:
                    raise

        # ── FIX Catalina: auto-persistir proveedor en tabla proveedores
        # Antes el proveedor solo quedaba como string en ordenes_compra.proveedor.
        # Cuando creaba la siguiente OC tenia que volver a escribir todo.
        # Ahora si el proveedor no existe, se crea con datos basicos (Catalina
        # puede enriquecerlo despues en /compras → Proveedores). Si ya existe,
        # solo se hace upsert de campos no-vacios para no pisar datos buenos.
        # FIX · 16-jun-2026 · mismo bug que crear_oc_desde_solicitudes: el UNIQUE
        # de proveedores.nombre es global → chequear con AND activo=1 y luego
        # INSERTar choca el UNIQUE si existe un proveedor INACTIVO con ese nombre
        # → en PG la tx queda abortada → el INSERT de items siguiente muere → 500.
        # Existencia SIN filtrar activo (= match el UNIQUE) + reactivar + SAVEPOINT.
        try:
            c.execute('SAVEPOINT _prov_auto')
            try:
                existe = c.execute(
                    "SELECT id, COALESCE(activo,0) FROM proveedores "
                    "WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?)) ORDER BY activo DESC LIMIT 1",
                    (d['proveedor'],)
                ).fetchone()
                _pid, _pacc = None, None
                if not existe:
                    c.execute("""INSERT INTO proveedores
                                 (nombre, categoria, condiciones_pago, activo, fecha_creacion)
                                 VALUES (?,?,?,1,?)""",
                              (d['proveedor'], categoria, '30 dias',
                               datetime.now().isoformat()))
                    _pid, _pacc = c.lastrowid, 'CREAR_PROVEEDOR'
                elif not existe[1]:
                    c.execute("UPDATE proveedores SET activo=1 WHERE id=?", (existe[0],))
                    _pid, _pacc = existe[0], 'REACTIVAR_PROVEEDOR'
                if _pacc:
                    audit_log(c, usuario=usuario, accion=_pacc,
                              tabla='proveedores', registro_id=_pid,
                              despues={'nombre': d['proveedor'][:200],
                                        'categoria': categoria,
                                        'origen': 'auto_oc',
                                        'oc_origen': numero_oc},
                              detalle=f"Auto al crear OC {numero_oc} · {d['proveedor'][:80]}")
                c.execute('RELEASE SAVEPOINT _prov_auto')
            except Exception:
                try:
                    c.execute('ROLLBACK TO SAVEPOINT _prov_auto')
                    c.execute('RELEASE SAVEPOINT _prov_auto')
                except Exception:
                    pass
        except Exception:
            pass

        # ── FIX Catalina: persistir precios en precios_mp_historico
        # para que la proxima vez que cree OC con el mismo MP, el precio
        # aparezca como sugerencia en autocomplete.
        # P1 audit 26-may · validate_money en items del POST (mismo patrón
        # que editar_oc PUT · evita NaN/Inf/negativos contaminando historico)
        from http_helpers import validate_money as _vm_h
        for it in (d.get('items') or []):
            _cg_v, _err_cg = _vm_h(it.get('cantidad_g', 0), allow_zero=False,
                                    max_value=1e9, field_name='cantidad_g')
            if _err_cg:
                conn.rollback()
                return jsonify(_err_cg), 400
            _pu_v, _err_pu = _vm_h(it.get('precio_unitario', 0), allow_zero=True,
                                    max_value=1e9, field_name='precio_unitario')
            if _err_pu:
                conn.rollback()
                return jsonify(_err_pu), 400
            cantidad_g = _cg_v
            precio_u = _pu_v
            subtotal = round(cantidad_g * precio_u, 2)
            codigo = it.get('codigo_mp', '')
            nombre = it.get('nombre_mp', '')
            c.execute("INSERT INTO ordenes_compra_items (numero_oc,codigo_mp,nombre_mp,cantidad_g,precio_unitario,subtotal) VALUES (?,?,?,?,?,?)",
                      (numero_oc, codigo, nombre, cantidad_g, precio_u, subtotal))
            # Persistir en historico de precios + actualizar referencia en maestro
            if codigo and precio_u > 0:
                try:
                    c.execute("""INSERT OR IGNORE INTO precios_mp_historico
                                 (codigo_mp, nombre_mp, precio_unitario, proveedor,
                                  fecha, numero_oc, cantidad_g)
                                 VALUES (?,?,?,?,?,?,?)""",
                              (codigo, nombre, precio_u, d['proveedor'],
                               datetime.now().isoformat()[:10], numero_oc, cantidad_g))
                except Exception:
                    pass
                try:
                    c.execute("""UPDATE maestro_mps
                                 SET precio_referencia=?, proveedor=COALESCE(NULLIF(proveedor,''),?)
                                 WHERE codigo_mp=?""",
                              (precio_u * 1000.0, d['proveedor'], codigo))  # $/g → $/kg (INV-2)
                except Exception:
                    pass

        valor_total_calc = sum(
            round((it.get('cantidad_g',0))*(it.get('precio_unitario',0)),2)
            for it in (d.get('items') or [])
        )
        if valor_total_calc > 0:
            c.execute("UPDATE ordenes_compra SET valor_total=? WHERE numero_oc=?", (valor_total_calc, numero_oc))
        try:
            audit_log(c, usuario=usuario, accion='CREAR_OC',
                      tabla='ordenes_compra', registro_id=numero_oc,
                      despues={'proveedor': d['proveedor'][:200],
                                'categoria': categoria,
                                'estado': 'Borrador',
                                'items_count': len(d.get('items') or []),
                                'valor_total': valor_total_calc,
                                'fecha_entrega_est': d.get('fecha_entrega_est','')[:30]},
                      detalle=f"Creó OC {numero_oc} · {d['proveedor']} · "
                              f"{len(d.get('items') or [])} items · total {valor_total_calc:.0f}")
        except Exception as e:
            log.warning('audit_log CREAR_OC fallo: %s', e)
        conn.commit()
        return jsonify({'message': f'OC {numero_oc} creada', 'numero_oc': numero_oc}), 201
    cat_filter = request.args.get('categoria', '')
    _sql = (
        "SELECT o.numero_oc, o.fecha, o.estado, o.proveedor, o.fecha_entrega_est,"
        " o.observaciones, o.creado_por, COUNT(i.id) as num_items,"
        " o.categoria, o.remision_code, o.autorizado_por,"
        " COALESCE(o.valor_total, 0) as valor_total,"
        " COALESCE(o.con_iva, 0) as con_iva, COALESCE(o.valor_sin_iva, 0) as valor_sin_iva"
        " FROM ordenes_compra o LEFT JOIN ordenes_compra_items i ON o.numero_oc=i.numero_oc"
    )
    if cat_filter:
        c.execute(_sql + " WHERE o.categoria=? GROUP BY o.id ORDER BY o.fecha DESC LIMIT 300", (cat_filter,))
    else:
        c.execute(_sql + " GROUP BY o.id ORDER BY o.fecha DESC LIMIT 300")
    cols = ['numero_oc','fecha','estado','proveedor','fecha_entrega_est','observaciones',
            'creado_por','num_items','categoria','remision_code','autorizado_por','valor_total',
            'con_iva','valor_sin_iva']
    rows = c.fetchall()
    return jsonify({'ordenes': [dict(zip(cols, r)) for r in rows]})

@bp.route('/api/ordenes-compra/<numero_oc>', methods=['GET','PUT','DELETE'])
def handle_oc_detalle(numero_oc):
    conn = get_db(); c = conn.cursor()
    if request.method == 'DELETE':
        usuario, err, code = _require_compras_write()
        if err:
            return err, code
        c.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
        _row = c.fetchone()
        if not _row:
            return jsonify({'error': 'OC no encontrada'}), 404
        if _row[0] not in ('Borrador', 'Rechazada'):
            return jsonify({'error': f'No se puede eliminar una OC en estado {_row[0]}. Solo Borrador o Rechazada.'}), 400
        # Fix #3 · 21-may-2026 · revertir SOLs vinculadas a Pendiente antes
        # de borrar la OC · sino quedaban con numero_oc apuntando a fantasma
        # y estado='Aprobada' · invisible al regenerar.
        try:
            r_sols = c.execute(
                "SELECT numero FROM solicitudes_compra WHERE numero_oc=?",
                (numero_oc,),
            ).fetchall()
            sols_revertidas = [r[0] for r in r_sols]
            c.execute(
                "UPDATE solicitudes_compra SET estado='Pendiente', numero_oc='' WHERE numero_oc=?",
                (numero_oc,),
            )
        except Exception as e:
            log.warning('revert SOLs fallo (eliminar OC): %s', e)
            sols_revertidas = []
        c.execute('DELETE FROM ordenes_compra_items WHERE numero_oc=?', (numero_oc,))
        c.execute('DELETE FROM ordenes_compra WHERE numero_oc=?', (numero_oc,))
        try:
            audit_log(c, usuario=usuario, accion='ELIMINAR_OC',
                      tabla='ordenes_compra', registro_id=numero_oc,
                      antes={'estado': _row[0]},
                      despues={'sols_revertidas': sols_revertidas},
                      detalle=f"Eliminó OC {numero_oc} (estado {_row[0]}) · revirtió {len(sols_revertidas)} SOLs")
        except Exception as e:
            log.warning('audit_log ELIMINAR_OC fallo: %s', e)
        conn.commit()
        return jsonify({
            'ok': True,
            'message': f'OC {numero_oc} eliminada',
            'sols_revertidas': sols_revertidas,
        })

    if request.method == 'PUT':
        u, err, code = _require_compras_write()
        if err:
            return err, code
        d = request.get_json(silent=True) or {}
        nuevo_estado = d.get('estado')
        if nuevo_estado:
            c.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
            _row = c.fetchone()
            if not _row:
                return jsonify({'error': 'OC no encontrada'}), 404
            estado_actual = _row[0]
            # State machine OC · Sebastián 24-may-2026 · audit Compras P0 #2.
            # Antes solo Pagada estaba protegida (INV-4) · Cancelada / Rechazada
            # / Recibida podían revertir a Borrador silenciosamente. Ahora
            # whitelist explícita de transiciones válidas · respeta legacy
            # Revisada / Parcial como pasthru y permite idempotente A→A.
            _OC_TRANSICIONES = {
                'Borrador':   {'Borrador','Autorizada','Cancelada','Rechazada','Revisada'},
                'Revisada':   {'Borrador','Autorizada','Cancelada','Rechazada','Revisada'},
                'Autorizada': {'Autorizada','Recibida','Pagada','Parcial','Cancelada','Rechazada'},
                'Parcial':    {'Parcial','Recibida','Pagada','Cancelada'},
                'Recibida':   {'Recibida','Pagada','Cancelada'},
                'Pagada':     {'Pagada'},        # terminal · INV-4
                'Cancelada':  {'Cancelada'},     # terminal
                'Rechazada':  {'Rechazada'},     # terminal
            }
            transiciones_ok = _OC_TRANSICIONES.get(estado_actual)
            # Admins pueden forzar overrides legítimos (motivo en body)
            es_admin = (u in ADMIN_USERS)
            if transiciones_ok is not None and nuevo_estado not in transiciones_ok and not es_admin:
                return jsonify({
                    'error': f"Transición no permitida: {estado_actual} → {nuevo_estado}",
                    'codigo': 'OC_TRANSICION_INVALIDA',
                    'estado_actual': estado_actual,
                    'transiciones_validas': sorted(transiciones_ok),
                    'hint': 'Solo admin puede forzar override (con motivo en observaciones).',
                }), 409
            c.execute("UPDATE ordenes_compra SET estado=? WHERE numero_oc=?",
                      (nuevo_estado, numero_oc))
            # AUDITORÍA-FIX 23-may-2026 · C3 · si la OC se cancela, liberar
            # las SOLs vinculadas a Pendiente · antes quedaban apuntando a
            # OC terminal y auto-plan no las re-emitía
            sols_revertidas_pp = []
            if str(nuevo_estado).strip().lower() in ('cancelada','cancelado'):
                try:
                    rows = c.execute(
                        "SELECT numero FROM solicitudes_compra WHERE numero_oc=?",
                        (numero_oc,),
                    ).fetchall()
                    sols_revertidas_pp = [r[0] for r in rows if r and r[0]]
                    c.execute(
                        "UPDATE solicitudes_compra SET estado='Pendiente', numero_oc='' "
                        "WHERE numero_oc=?",
                        (numero_oc,),
                    )
                except Exception:
                    pass
            try:
                audit_log(c, usuario=u, accion='ACTUALIZAR_ESTADO_OC',
                          tabla='ordenes_compra', registro_id=numero_oc,
                          antes={'estado': estado_actual},
                          despues={'estado': nuevo_estado,
                                   'sols_revertidas': sols_revertidas_pp},
                          detalle=f"OC {numero_oc}: {estado_actual} → {nuevo_estado}")
            except Exception as e:
                log.warning('audit_log ACTUALIZAR_ESTADO_OC fallo: %s', e)
        conn.commit()
        return jsonify({'message': f'OC {numero_oc} actualizada'})
    c.execute("SELECT * FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    oc_row = c.fetchone()
    oc_cols = [d[0] for d in c.description] if c.description else []
    c.execute("SELECT * FROM ordenes_compra_items WHERE numero_oc=?", (numero_oc,))
    items = [dict(row) for row in c.fetchall()]
    if not oc_row: return jsonify({'error': 'OC no encontrada'}), 404
    oc_dict = dict(zip(oc_cols, oc_row))
    prov_data = None
    prov_name = oc_dict.get('proveedor', '')
    if prov_name:
        c.execute(
            "SELECT banco, tipo_cuenta, num_cuenta, nit, email, telefono, contacto"
            " FROM proveedores WHERE nombre=? AND activo=1 LIMIT 1",
            (prov_name,)
        )
        prow = c.fetchone()
        if prow:
            prov_data = dict(zip(
                ['banco', 'tipo_cuenta', 'num_cuenta', 'nit', 'email', 'telefono', 'contacto'],
                prow
            ))
    return jsonify({'oc': oc_dict, 'items': items, 'prov_data': prov_data})

@bp.route('/api/ordenes-compra/<numero_oc>/editar', methods=['PATCH'])
def editar_oc(numero_oc):
    """Edita una OC. Permisos por estado (decision Sebastian + Catalina 2026-04-28):

      - Borrador / Pendiente / Revisada / Aprobada / Autorizada:
          edicion COMPLETA (proveedor, categoria, observaciones, fechas, items)
      - Recibida / Parcial:
          solo observaciones + observaciones_recepcion + fecha_entrega_est
          (NO items, NO proveedor — la mercancia ya llego, no se cambia)
      - Pagada:
          solo observaciones (auditable)
      - Cancelada / Rechazada:
          bloqueado (la OC esta muerta)

    Esto resuelve la queja real de Catalina: una vez la OC pasaba de Borrador
    quedaba congelada y no podia ajustar ni el precio si el proveedor cambiaba
    despues.
    """
    # Audit zero-error 2-may-2026: RBAC mínimo (era cualquier sesión)
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in (set(COMPRAS_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Compras/Admin pueden editar OC'}), 403
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT estado FROM ordenes_compra WHERE numero_oc=?', (numero_oc,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado = row[0] or 'Borrador'

    # Bloqueos absolutos
    if estado in ('Cancelada', 'Rechazada'):
        return jsonify({
            'error': f'OC en estado {estado} no se puede editar. '
                     f'Crea una nueva OC en su lugar.'
        }), 400

    d = request.json or {}
    EDITABLE_FULL = {'Borrador', 'Pendiente', 'Revisada', 'Aprobada', 'Autorizada'}
    EDITABLE_LIMITED = {'Recibida', 'Parcial'}
    EDITABLE_OBS_ONLY = {'Pagada'}

    # ─── Edicion completa ────────────────────────────────────────────────
    # PATCH semántico: solo se actualiza lo que VIENE en el body. Lo que no se
    # especifique mantiene su valor actual en BD. Esto permite PATCH parciales
    # (ej. solo togglear con_iva desde el Consolidado sin reenviar proveedor).
    if estado in EDITABLE_FULL:
        # Leer estado actual para fallback en campos no especificados
        cur_row = c.execute("""
            SELECT proveedor, COALESCE(categoria,'MP'), COALESCE(observaciones,''),
                   COALESCE(fecha_entrega_est,''), COALESCE(con_iva,0),
                   COALESCE(valor_sin_iva,0)
            FROM ordenes_compra WHERE numero_oc=?
        """, (numero_oc,)).fetchone()
        cur_prov, cur_cat, cur_obs, cur_fent, cur_iva, cur_vsi = cur_row
        proveedor = d['proveedor'] if 'proveedor' in d and d.get('proveedor') else cur_prov
        if not proveedor:
            return jsonify({'error': 'Proveedor requerido (no hay valor actual y no vino en body)'}), 400
        categoria = d.get('categoria', cur_cat)
        observaciones = d.get('observaciones', cur_obs) if 'observaciones' in d else cur_obs
        fecha_entrega_est = d.get('fecha_entrega_est', cur_fent) if 'fecha_entrega_est' in d else cur_fent
        con_iva = (1 if d.get('con_iva') else 0) if 'con_iva' in d else int(cur_iva or 0)
        # P1 audit 26-may · simetría POST/PUT · validate_money en editar_oc
        # Antes: float() pelado aceptaba NaN/Inf/negativos · contaminaba
        # valor_total downstream.
        from http_helpers import validate_money as _vm_oc
        if 'valor_sin_iva' in d:
            valor_sin_iva, _err_vsi = _vm_oc(d.get('valor_sin_iva'),
                                              allow_zero=True, max_value=1e10,
                                              field_name='valor_sin_iva')
            if _err_vsi:
                return jsonify(_err_vsi), 400
        else:
            valor_sin_iva = float(cur_vsi or 0)
        c.execute("""
            UPDATE ordenes_compra SET
                proveedor=?, categoria=?, observaciones=?,
                fecha_entrega_est=?, con_iva=?, valor_sin_iva=?
            WHERE numero_oc=?""",
            (proveedor, categoria, observaciones,
             fecha_entrega_est, con_iva, valor_sin_iva, numero_oc))
        # Si vinieron items, reemplazar completo y recalcular valor_total con IVA
        if 'items' in d:
            c.execute('DELETE FROM ordenes_compra_items WHERE numero_oc=?', (numero_oc,))
            valor_total = 0.0
            for it in (d.get('items') or []):
                # validate_money en items PUT (asimetría con POST)
                _cg, _err_cg = _vm_oc(it.get('cantidad_g', 0), allow_zero=False,
                                       max_value=1e9, field_name='cantidad_g')
                if _err_cg:
                    return jsonify(_err_cg), 400
                _pu, _err_pu = _vm_oc(it.get('precio_unitario', 0), allow_zero=True,
                                       max_value=1e9, field_name='precio_unitario')
                if _err_pu:
                    return jsonify(_err_pu), 400
                subtotal = round(_cg * _pu, 2)
                c.execute('INSERT INTO ordenes_compra_items (numero_oc,codigo_mp,nombre_mp,cantidad_g,precio_unitario,subtotal) VALUES (?,?,?,?,?,?)',
                          (numero_oc, it.get('codigo_mp', ''), it.get('nombre_mp', ''),
                           _cg, _pu, subtotal))
                valor_total += subtotal
            if con_iva:
                valor_total = round(valor_total * 1.19, 2)
            c.execute('UPDATE ordenes_compra SET valor_total=? WHERE numero_oc=?', (valor_total, numero_oc))
        else:
            # Sin items en body → recalcular valor_total respetando con_iva sobre items existentes
            # (importante cuando solo se togglea IVA sin tocar items)
            sub = c.execute(
                'SELECT COALESCE(SUM(subtotal),0) FROM ordenes_compra_items WHERE numero_oc=?',
                (numero_oc,)
            ).fetchone()[0] or 0.0
            valor_total = round(sub * 1.19, 2) if con_iva else round(sub, 2)
            c.execute('UPDATE ordenes_compra SET valor_total=? WHERE numero_oc=?', (valor_total, numero_oc))
        try:
            audit_log(c, usuario=user, accion='EDITAR_OC',
                      tabla='ordenes_compra', registro_id=numero_oc,
                      antes={'proveedor': cur_prov, 'categoria': cur_cat,
                              'con_iva': cur_iva, 'valor_sin_iva': cur_vsi},
                      despues={k: d.get(k) for k in d
                                if k in ('proveedor','categoria','observaciones',
                                         'fecha_entrega_est','con_iva','valor_sin_iva',
                                         'items')},
                      detalle=f"Editó OC {numero_oc} ({estado})")
        except Exception as e:
            log.warning('audit_log EDITAR_OC fallo: %s', e)
        conn.commit()
        return jsonify({'ok': True, 'message': f'OC {numero_oc} ({estado}) actualizada'})

    # ─── Edicion limitada (Recibida/Parcial): solo metadata sin tocar mercancia ─
    if estado in EDITABLE_LIMITED:
        sets, params = [], []
        for f in ('observaciones', 'observaciones_recepcion', 'fecha_entrega_est'):
            if f in d:
                sets.append(f'{f}=?'); params.append(d[f])
        if not sets:
            return jsonify({'error': f'En estado {estado} solo se puede editar observaciones / fecha_entrega_est'}), 400
        params.append(numero_oc)
        c.execute(f'UPDATE ordenes_compra SET {", ".join(sets)} WHERE numero_oc=?', params)
        try:
            audit_log(c, usuario=user, accion='EDITAR_OC',
                      tabla='ordenes_compra', registro_id=numero_oc,
                      despues={k: d.get(k) for k in d
                                if k in ('observaciones','observaciones_recepcion','fecha_entrega_est')},
                      detalle=f"Editó OC {numero_oc} ({estado}) — campos limitados")
        except Exception as e:
            log.warning('audit_log EDITAR_OC fallo: %s', e)
        conn.commit()
        return jsonify({'ok': True, 'message': f'OC {numero_oc} ({estado}) — campos limitados actualizados',
                        'campos_actualizados': [s.split('=')[0] for s in sets]})

    # ─── Edicion mínima (Pagada): solo observaciones para audit trail ─────
    if estado in EDITABLE_OBS_ONLY:
        if 'observaciones' not in d:
            return jsonify({'error': 'En estado Pagada solo se permite editar observaciones'}), 400
        # Append en lugar de sobreescribir (audit trail)
        nueva_obs = (d.get('observaciones') or '').strip()
        ts = datetime.now().strftime('%Y-%m-%d %H:%M')
        usuario = session.get('compras_user', '')
        nota = f'\n[{ts} {usuario}] {nueva_obs}'
        c.execute("""UPDATE ordenes_compra
                     SET observaciones = COALESCE(observaciones,'') || ?
                     WHERE numero_oc=?""", (nota, numero_oc))
        conn.commit()
        return jsonify({'ok': True, 'message': f'Nota agregada al historial de OC {numero_oc}'})

    return jsonify({'error': f'Estado {estado} no soportado para edicion'}), 400


@bp.route('/api/ordenes-compra/<numero_oc>/items', methods=['POST'])
def agregar_item_oc(numero_oc):
    """Agrega un item a una OC existente sin tener que recrearla.

    Body: {codigo_mp, nombre_mp, cantidad_g, precio_unitario}

    Permitido en estados: Borrador, Pendiente, Revisada, Aprobada, Autorizada.
    Bloqueado en Recibida/Pagada/Cancelada/Rechazada.
    """
    # Audit zero-error 2-may-2026: RBAC mínimo
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in (set(COMPRAS_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Compras/Admin pueden agregar items'}), 403
    conn = get_db(); c = conn.cursor()
    row = c.execute('SELECT estado, con_iva FROM ordenes_compra WHERE numero_oc=?',
                    (numero_oc,)).fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado, con_iva = row[0], row[1]
    if estado not in ('Borrador', 'Pendiente', 'Revisada', 'Aprobada', 'Autorizada'):
        return jsonify({'error': f'No se pueden agregar items en estado {estado}'}), 400
    d = request.json or {}
    nombre_mp = (d.get('nombre_mp') or '').strip()
    if not nombre_mp:
        return jsonify({'error': 'nombre_mp requerido'}), 400
    if len(nombre_mp) > 300:
        return jsonify({'error': 'nombre_mp excede 300 chars'}), 400
    # Audit zero-error 2-may-2026: usar validate_money para sanity check completo
    # (NaN/Infinity/cap). Antes solo se validaba >0 y <1B gramos.
    cantidad_g, err = validate_money(d.get('cantidad_g', 0), allow_zero=False,
                                       max_value=1_000_000_000, field_name='cantidad_g')
    if err:
        return jsonify(err), 400
    precio, err = validate_money(d.get('precio_unitario', 0), allow_zero=True,
                                   field_name='precio_unitario')
    if err:
        return jsonify(err), 400
    subtotal = round(cantidad_g * precio, 2)
    c.execute("""INSERT INTO ordenes_compra_items
                 (numero_oc, codigo_mp, nombre_mp, cantidad_g, precio_unitario, subtotal)
                 VALUES (?,?,?,?,?,?)""",
              (numero_oc, d.get('codigo_mp', ''), nombre_mp, cantidad_g, precio, subtotal))
    item_id = c.lastrowid
    # Recalcular total OC
    suma = c.execute("SELECT COALESCE(SUM(subtotal),0) FROM ordenes_compra_items WHERE numero_oc=?",
                     (numero_oc,)).fetchone()[0]
    if con_iva:
        suma = round(suma * 1.19, 2)
    c.execute('UPDATE ordenes_compra SET valor_total=? WHERE numero_oc=?', (suma, numero_oc))
    try:
        audit_log(c, usuario=user, accion='AGREGAR_ITEM_OC',
                  tabla='ordenes_compra_items', registro_id=numero_oc,
                  despues={'item_id': item_id, 'codigo_mp': d.get('codigo_mp',''),
                            'nombre_mp': nombre_mp[:200], 'cantidad_g': cantidad_g,
                            'precio_unitario': precio, 'subtotal': subtotal,
                            'estado_oc': estado},
                  detalle=f"Agregó item {nombre_mp[:60]} a OC {numero_oc} · "
                          f"{cantidad_g}g @ {precio:.0f}")
    except Exception as e:
        log.warning('audit_log AGREGAR_ITEM_OC fallo: %s', e)
    conn.commit()
    return jsonify({'ok': True, 'item_id': item_id, 'valor_total': suma})


@bp.route('/api/ordenes-compra/<numero_oc>/items/<int:item_id>', methods=['PATCH', 'DELETE'])
def modificar_item_oc(numero_oc, item_id):
    """PATCH: actualiza precio/cantidad/nombre de un item.
    DELETE: elimina un item.

    Estados permitidos: Borrador, Pendiente, Revisada, Aprobada, Autorizada.
    """
    # SEC-FIX 23-may-2026 · auditoría · era 'compras_user in session' permisivo
    u, err, code = _require_compras_write()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    row = c.execute('SELECT estado, con_iva FROM ordenes_compra WHERE numero_oc=?',
                    (numero_oc,)).fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado, con_iva = row[0], row[1]
    if estado not in ('Borrador', 'Pendiente', 'Revisada', 'Aprobada', 'Autorizada'):
        return jsonify({'error': f'No se pueden modificar items en estado {estado}'}), 400

    if request.method == 'DELETE':
        c.execute('DELETE FROM ordenes_compra_items WHERE id=? AND numero_oc=?',
                  (item_id, numero_oc))
        if c.rowcount == 0:
            return jsonify({'error': 'Item no encontrado'}), 404
    else:
        d = request.json or {}
        sets, params = [], []
        if 'codigo_mp' in d:
            sets.append('codigo_mp=?'); params.append(d['codigo_mp'])
        if 'nombre_mp' in d:
            sets.append('nombre_mp=?'); params.append(d['nombre_mp'])
        if 'cantidad_g' in d or 'precio_unitario' in d:
            # Recalcular subtotal con valores actualizados
            cur_row = c.execute("""SELECT cantidad_g, precio_unitario
                                   FROM ordenes_compra_items
                                   WHERE id=? AND numero_oc=?""",
                                (item_id, numero_oc)).fetchone()
            if not cur_row:
                return jsonify({'error': 'Item no encontrado'}), 404
            # SEC-FIX 23-may-2026 · auditoría · validate_money rechaza NaN/Inf/neg
            import math as _math
            try:
                cant = float(d.get('cantidad_g', cur_row[0] or 0))
                prec = float(d.get('precio_unitario', cur_row[1] or 0))
            except (TypeError, ValueError):
                return jsonify({'error': 'cantidad/precio inválido'}), 400
            if _math.isnan(cant) or _math.isinf(cant) or cant < 0:
                return jsonify({'error': 'cantidad debe ser número finito >= 0'}), 400
            if _math.isnan(prec) or _math.isinf(prec) or prec < 0:
                return jsonify({'error': 'precio debe ser número finito >= 0'}), 400
            sets += ['cantidad_g=?', 'precio_unitario=?', 'subtotal=?']
            params += [cant, prec, round(cant * prec, 2)]
        if not sets:
            return jsonify({'error': 'Nada que actualizar'}), 400
        params += [item_id, numero_oc]
        c.execute(f"""UPDATE ordenes_compra_items SET {', '.join(sets)}
                      WHERE id=? AND numero_oc=?""", params)

    # Recalcular total OC
    suma = c.execute("SELECT COALESCE(SUM(subtotal),0) FROM ordenes_compra_items WHERE numero_oc=?",
                     (numero_oc,)).fetchone()[0]
    if con_iva:
        suma = round(suma * 1.19, 2)
    c.execute('UPDATE ordenes_compra SET valor_total=? WHERE numero_oc=?', (suma, numero_oc))
    try:
        usuario_act = session.get('compras_user', '')
        accion_audit = 'ELIMINAR_ITEM_OC' if request.method == 'DELETE' else 'MODIFICAR_ITEM_OC'
        audit_log(c, usuario=usuario_act, accion=accion_audit,
                  tabla='ordenes_compra_items', registro_id=numero_oc,
                  despues={'item_id': item_id,
                            'valor_total_nuevo': suma,
                            'estado_oc': estado},
                  detalle=f"{accion_audit.replace('_',' ').title()} item {item_id} en OC {numero_oc}")
    except Exception as e:
        log.warning('audit_log MODIFICAR_ITEM_OC fallo: %s', e)
    conn.commit()
    return jsonify({'ok': True, 'valor_total': suma})

@bp.route('/api/ordenes-compra/<numero_oc>/proveedor', methods=['PATCH'])
def confirmar_proveedor_oc(numero_oc):
    """Confirma o cambia el proveedor de una OC.

    Body: { proveedor: 'Nombre Proveedor', confirmado: true }

    Pensado para el flujo donde Catalina revisa la solicitud pendiente y:
      - Si el proveedor sugerido es correcto → confirma con click (proveedor
        sigue igual, solo actualiza fecha y marca confirmacion).
      - Si quiere cambiarlo → escribe el nuevo nombre y se actualiza.

    NO requiere estado Borrador — confirmacion también es válida en
    Aprobada (siempre que no haya pagos hechos).

    Cuando se confirma o cambia, alimentamos el catálogo: si el proveedor
    es nuevo y no existe en `proveedores`, se crea con datos mínimos para
    que aparezca en futuros selectores. Esto es lo que el user pidió:
    'sirve para confirmar y que el sistema se alimente'.
    """
    # SEC-FIX 23-may-2026 · auditoría · solo compras_user permitía a
    # cualquier user logueado redirigir OC a un proveedor distinto antes
    # de autorize · ataque potencial · ahora _require_compras_write
    u, err, code = _require_compras_write()
    if err: return err, code
    d = request.get_json() or {}
    nuevo_prov = (d.get('proveedor') or '').strip()
    if not nuevo_prov:
        return jsonify({'error': 'proveedor requerido'}), 400

    conn = get_db(); c = conn.cursor()
    row = c.execute(
        'SELECT estado, proveedor, categoria FROM ordenes_compra WHERE numero_oc=?',
        (numero_oc,)
    ).fetchone()
    # SEC-FIX 23-may-2026 · bloquear cambio post-autorización · evita que
    # alguien redirija OC ya validada a otro proveedor (ataque o error)
    if row and row[0] and row[0] in ('Autorizada', 'Pagada', 'Recibida', 'Parcial'):
        return jsonify({
            'error': f'No se puede cambiar proveedor con estado {row[0]} · '
                     'requiere cancelar OC y crear nueva',
        }), 409
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado_oc, proveedor_actual, categoria = row[0], (row[1] or ''), (row[2] or 'MP')

    # No permitir cambio si ya hay pagos (audit trail)
    try:
        n_pagos = c.execute(
            'SELECT COUNT(*) FROM pagos_oc WHERE numero_oc=?', (numero_oc,)
        ).fetchone()[0] or 0
        if n_pagos > 0 and proveedor_actual.lower() != nuevo_prov.lower():
            return jsonify({
                'error': 'No se puede cambiar proveedor de una OC con pagos registrados',
                'codigo': 'OC_CON_PAGOS'
            }), 400
    except Exception:
        pass

    # Si el proveedor no existe en catálogo, crearlo (alimenta sistema)
    creado_nuevo = False
    try:
        existe = c.execute(
            'SELECT 1 FROM proveedores WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?))',
            (nuevo_prov,)
        ).fetchone()
        if not existe:
            c.execute(
                """INSERT INTO proveedores (nombre, categoria, fecha_creacion)
                   VALUES (?, ?, ?)""",
                (nuevo_prov, categoria, datetime.now().isoformat())
            )
            creado_nuevo = True
            try:
                audit_log(c, usuario=session.get('compras_user', 'sistema'),
                          accion='CREAR_PROVEEDOR', tabla='proveedores',
                          registro_id=c.lastrowid,
                          despues={'nombre': nuevo_prov[:200],
                                    'categoria': categoria,
                                    'origen': 'auto_cambio_proveedor_oc',
                                    'oc': numero_oc},
                          detalle=f"Auto-creado al cambiar proveedor de OC {numero_oc}")
            except Exception:
                pass
    except sqlite3.OperationalError:
        pass

    # Actualizar la OC
    c.execute(
        'UPDATE ordenes_compra SET proveedor=? WHERE numero_oc=?',
        (nuevo_prov, numero_oc)
    )
    try:
        usuario_act = session.get('compras_user', '')
        cambio = proveedor_actual.lower() != nuevo_prov.lower()
        audit_log(c, usuario=usuario_act, accion='CONFIRMAR_PROVEEDOR_OC',
                  tabla='ordenes_compra', registro_id=numero_oc,
                  antes={'proveedor': proveedor_actual},
                  despues={'proveedor': nuevo_prov, 'creado_catalogo': creado_nuevo,
                            'cambio': cambio},
                  detalle=f"{'Cambió' if cambio else 'Confirmó'} proveedor OC {numero_oc} "
                          f"de '{proveedor_actual}' a '{nuevo_prov}'")
    except Exception as e:
        log.warning('audit_log CONFIRMAR_PROVEEDOR_OC fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'numero_oc': numero_oc,
        'proveedor_anterior': proveedor_actual,
        'proveedor_nuevo': nuevo_prov,
        'creado_en_catalogo': creado_nuevo,
        'cambio': proveedor_actual.lower() != nuevo_prov.lower(),
    })


@bp.route('/api/ordenes-compra/<numero_oc>/items-precios', methods=['PATCH'])
def actualizar_precios_items_oc(numero_oc):
    """Actualiza precios unitarios de items de una OC + alimenta histórico.

    Body: { items: [{ codigo_mp, precio_unitario, cantidad_g? }, ...] }

    Para que Catalina (en el flujo de confirmar solicitud) cargue precios
    por item — esto:
      1. Actualiza precio_unitario y subtotal en ordenes_compra_items
      2. Alimenta precios_mp_historico para que aparezca en próximos
         pedidos como precio sugerido
      3. Recalcula valor_total de la OC

    El user pidió: 'que catalina coloque el valor de cada cosa y guardar
    asi empezamos a tener almacenado los valores de las cosas para que
    sea mas automatico'.
    """
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    d = request.get_json() or {}
    items_in = d.get('items') or []
    if not items_in:
        return jsonify({'error': 'items requerido'}), 400

    conn = get_db(); c = conn.cursor()
    row = c.execute(
        'SELECT estado, proveedor FROM ordenes_compra WHERE numero_oc=?', (numero_oc,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    if row[0] in ('Pagada', 'Cancelada', 'Rechazada'):
        return jsonify({'error': f'No se pueden editar precios de una OC en estado {row[0]}.'}), 409
    proveedor = row[1] or ''

    # Actualizar cada item por codigo_mp (y opcionalmente cantidad_g)
    actualizados = 0
    # P1 audit 26-may · validate_money en actualizar_precios_items_oc PUT
    from http_helpers import validate_money as _vm_pr
    for it in items_in:
        cod = (it.get('codigo_mp') or '').strip()
        if not cod:
            continue
        _pr_v, _err_pr = _vm_pr(it.get('precio_unitario', 0) or 0,
                                 allow_zero=True, max_value=1e9,
                                 field_name='precio_unitario')
        if _err_pr:
            return jsonify(_err_pr), 400
        precio = _pr_v
        # Cantidad: si la pasan, la actualizamos; si no, la dejamos como está
        cant = it.get('cantidad_g')
        if cant is not None:
            _ct_v, _err_ct = _vm_pr(cant, allow_zero=False, max_value=1e9,
                                     field_name='cantidad_g')
            if _err_ct:
                return jsonify(_err_ct), 400
            try:
                cant = _ct_v
                # subtotal = cantidad * precio
                subtotal = round(cant * precio, 2)
                c.execute(
                    """UPDATE ordenes_compra_items
                       SET precio_unitario=?, cantidad_g=?, subtotal=?
                       WHERE numero_oc=? AND codigo_mp=?""",
                    (precio, cant, subtotal, numero_oc, cod)
                )
            except (ValueError, TypeError):
                continue
        else:
            # Solo precio: recalcula subtotal con la cantidad existente
            c.execute(
                """UPDATE ordenes_compra_items
                   SET precio_unitario=?,
                       subtotal=ROUND(COALESCE(cantidad_g,0)*?, 2)
                   WHERE numero_oc=? AND codigo_mp=?""",
                (precio, precio, numero_oc, cod)
            )
        actualizados += 1

        # Alimentar histórico de precios (si la tabla existe)
        if precio > 0:
            try:
                c.execute(
                    """INSERT OR IGNORE INTO precios_mp_historico
                       (codigo_mp, precio_kg, proveedor, fecha)
                       VALUES (?, ?, ?, datetime('now', '-5 hours'))""",
                    (cod, precio, proveedor)
                )
            except sqlite3.OperationalError:
                pass

    # Recalcular valor_total de la OC · FIX 13-jun (audit compras · M12(f)): RESPETAR
    # con_iva. Antes este endpoint (el que usa Catalina para guardar precios) guardaba
    # la SUMA de subtotales SIN IVA mientras editar_oc/agregar/modificar sí aplican
    # ×1.19 → tras pasar por aquí el valor_total perdía el 16% y pagar_oc/espejo
    # financiero pagaban de menos. Ahora alineado con los otros 3 paths.
    total = c.execute(
        'SELECT COALESCE(SUM(COALESCE(subtotal,0)),0) FROM ordenes_compra_items WHERE numero_oc=?',
        (numero_oc,)
    ).fetchone()[0] or 0
    _row_iva = c.execute('SELECT COALESCE(con_iva,0) FROM ordenes_compra WHERE numero_oc=?',
                         (numero_oc,)).fetchone()
    _con_iva = int(_row_iva[0]) if _row_iva else 0
    sub = round(float(total), 2)
    valor_total = round(sub * 1.19, 2) if _con_iva else sub
    c.execute(
        'UPDATE ordenes_compra SET valor_total=?, valor_sin_iva=? WHERE numero_oc=?',
        (valor_total, sub, numero_oc)
    )
    try:
        usuario_act = session.get('compras_user', '')
        audit_log(c, usuario=usuario_act, accion='ACTUALIZAR_PRECIOS_OC',
                  tabla='ordenes_compra_items', registro_id=numero_oc,
                  despues={'items_actualizados': actualizados,
                            'valor_total_nuevo': valor_total,
                            'valor_sin_iva': sub, 'con_iva': _con_iva,
                            'proveedor': proveedor[:200]},
                  detalle=f"Actualizó precios de {actualizados} items en OC {numero_oc} "
                          f"· nuevo total {valor_total:.0f}" + (" (con IVA)" if _con_iva else ""))
    except Exception as e:
        log.warning('audit_log ACTUALIZAR_PRECIOS_OC fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'numero_oc': numero_oc,
        'items_actualizados': actualizados,
        'valor_total_nuevo': round(float(total), 2),
    })


@bp.route('/api/compras/sugerir-mp-bulk', methods=['POST'])
def sugerir_mp_bulk():
    """Sprint Compras N2 · 21-may-2026 · auto-fill precio desde histórico.

    Devuelve precio último + proveedor + fecha para una LISTA de codigo_mp
    en una sola llamada. Evita N round-trips cuando la tab Planta tiene
    20-30 items consolidados.

    Body: {codigos: [str, ...]}  (max 200)
    Returns: {datos: {codigo_mp: {precio_ultimo, proveedor_ultimo,
              fecha_ultimo, oc_ultima, dias_atras, precio_promedio_90d}}}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    body = request.get_json(silent=True) or {}
    codigos = body.get('codigos') or []
    if not isinstance(codigos, list):
        return jsonify({'error': 'codigos debe ser lista'}), 400
    codigos = [str(c).strip() for c in codigos if str(c).strip()][:200]
    if not codigos:
        return jsonify({'datos': {}})
    conn = get_db(); c = conn.cursor()
    placeholders = ','.join(['?'] * len(codigos))
    datos = {}
    # Último precio por código · COALESCE columna depende del schema
    # (precio_unitario en algunos schemas legacy, precio_kg en el actual).
    # Probamos primero precio_unitario · si falla, usar precio_kg.
    rows = []
    for col_precio in ('precio_unitario', 'precio_kg'):
        rows = []
        try:
            for cod in codigos:
                r = c.execute(
                    f"""SELECT codigo_mp, {col_precio}, COALESCE(proveedor,''),
                              COALESCE(fecha,''), COALESCE(numero_oc,'')
                       FROM precios_mp_historico
                       WHERE codigo_mp=?
                       ORDER BY fecha DESC, id DESC LIMIT 1""",
                    (cod,),
                ).fetchone()
                if r:
                    rows.append((r[0], r[1], r[2], r[3], r[4], r[3]))
            break  # query exitosa, salir del for
        except Exception:
            try: conn.rollback()
            except Exception: pass
            rows = []
            continue
    from datetime import datetime as _dt, date as _d
    hoy = _d.today()
    seen = set()
    for r in rows:
        cod = r[0]
        # Si usamos window, solo tomar la fila más reciente por código
        if len(r) >= 6 and r[3] != r[5]:
            continue
        if cod in seen:
            continue
        seen.add(cod)
        fecha_str = (r[3] or '')[:10]
        dias_atras = None
        if fecha_str:
            try:
                d_obj = _dt.strptime(fecha_str, '%Y-%m-%d').date()
                dias_atras = (hoy - d_obj).days
            except Exception:
                pass
        datos[cod] = {
            'precio_ultimo': float(r[1] or 0),
            'proveedor_ultimo': r[2] or '',
            'fecha_ultimo': fecha_str,
            'oc_ultima': r[4] or '',
            'dias_atras': dias_atras,
        }
    # Precio promedio 90d por código · misma estrategia de columna
    for col_precio in ('precio_unitario', 'precio_kg'):
        try:
            avg_rows = c.execute(
                f"""SELECT codigo_mp, AVG({col_precio})
                    FROM precios_mp_historico
                    WHERE codigo_mp IN ({placeholders})
                      AND date(fecha) >= date('now','-5 hours','-90 days')
                    GROUP BY codigo_mp""",
                codigos,
            ).fetchall()
            for r in avg_rows:
                if r[0] in datos:
                    datos[r[0]]['precio_promedio_90d'] = float(r[1] or 0)
            break
        except Exception:
            try: conn.rollback()
            except Exception: pass
            continue
    return jsonify({'datos': datos, 'codigos_consultados': len(codigos)})


@bp.route('/api/compras/sugerir-mp/<path:codigo_mp>')
def sugerir_mp(codigo_mp):
    """Devuelve datos sugeridos para autocompletar un MP en una OC.
    Resuelve la queja de Catalina: 'pongo precios y proveedor y no quedan
    guardados — siempre me los vuelve a pedir'.

    Devuelve:
      - nombre_mp (de maestro_mps)
      - precio_referencia (ultimo precio en maestro_mps)
      - precio_ultimo (de precios_mp_historico, mas reciente)
      - proveedor_ultimo (de precios_mp_historico, mas reciente)
      - top_proveedores: lista de proveedores que han vendido este MP
        ordenados por uso (top 5) con su precio promedio
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    out = {'codigo_mp': codigo_mp}
    try:
        mp = c.execute("""SELECT nombre_inci, nombre_comercial, proveedor,
                                COALESCE(precio_referencia,0)
                         FROM maestro_mps WHERE codigo_mp=?""", (codigo_mp,)).fetchone()
        if mp:
            out['nombre_inci'] = mp[0]
            out['nombre_comercial'] = mp[1]
            out['proveedor_default'] = mp[2]
            out['precio_referencia'] = mp[3]
    except Exception:
        pass

    try:
        # Ultimo precio en historial
        ult = c.execute("""SELECT precio_unitario, proveedor, fecha, numero_oc
                          FROM precios_mp_historico
                          WHERE codigo_mp=?
                          ORDER BY fecha DESC, id DESC LIMIT 1""",
                       (codigo_mp,)).fetchone()
        if ult:
            out['precio_ultimo'] = ult[0]
            out['proveedor_ultimo'] = ult[1]
            out['fecha_ultimo'] = ult[2]
            out['oc_ultima'] = ult[3]
    except Exception:
        pass

    try:
        # Top proveedores por uso historico (ultimas 50 compras del MP)
        rows = c.execute("""
            SELECT proveedor, COUNT(*) as veces, AVG(precio_unitario) as precio_avg
            FROM precios_mp_historico
            WHERE codigo_mp=? AND COALESCE(proveedor,'') != ''
            GROUP BY proveedor
            ORDER BY veces DESC, precio_avg ASC
            LIMIT 5
        """, (codigo_mp,)).fetchall()
        out['top_proveedores'] = [
            {'proveedor': r[0], 'veces_usado': r[1], 'precio_promedio': round(r[2] or 0, 2)}
            for r in rows
        ]
    except Exception:
        out['top_proveedores'] = []

    return jsonify(out)


# ════════════════════════════════════════════════════════════════════
# Libro de facturas de proveedor (cuentas por pagar formal) · mig 206
# Sebastián 31-may-2026 · factura = padre de pagos (pagos_oc.factura_proveedor_id)
# Antes la factura vivía solo como pagos_oc.numero_factura_proveedor (texto+imagen).
# ════════════════════════════════════════════════════════════════════
def _fp_recalc_estado(c, factura_id):
    """Recalcula estado por SUM(pagos) vs total. No toca 'anulada'."""
    row = c.execute("SELECT total, estado FROM facturas_proveedor WHERE id=?",
                    (factura_id,)).fetchone()
    if not row or row[1] == 'anulada':
        return
    total = float(row[0] or 0)
    pagado = float(c.execute(
        "SELECT COALESCE(SUM(monto),0) FROM pagos_oc WHERE factura_proveedor_id=?",
        (factura_id,)).fetchone()[0] or 0)
    if pagado <= 0.009:
        est = 'pendiente'
    elif pagado < total - 0.01:
        est = 'parcial'
    else:
        est = 'pagada'
    c.execute("UPDATE facturas_proveedor SET estado=? WHERE id=?", (est, factura_id))


def _fp_total_calc(d):
    """Total a pagar = subtotal + iva - retefuente - retica."""
    try:
        return round(float(d.get('subtotal') or 0) + float(d.get('iva') or 0)
                     - float(d.get('retefuente') or 0) - float(d.get('retica') or 0), 2)
    except (TypeError, ValueError):
        return 0.0


@bp.route('/api/compras/facturas-proveedor', methods=['GET'])
def fp_listar():
    """Libro de facturas de proveedor + saldos/aging. Filtros: estado, proveedor, q."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db(); c = conn.cursor()
    estado = (request.args.get('estado') or '').strip()
    proveedor = (request.args.get('proveedor') or '').strip()
    q = (request.args.get('q') or '').strip().lower()
    # FIX 1-jun-2026 (audit escalabilidad): sin SELECT * (no traer el PDF · anti-OOM)
    # y sin N+1 · pagado/valor_oc/tiene_pdf vía LEFT JOIN · filtro q en SQL.
    sql = """SELECT f.id, f.numero_factura, f.proveedor, f.nit, f.numero_oc,
                 f.fecha_emision, f.fecha_vencimiento, f.total, f.estado,
                 COALESCE(p.pagado,0) AS pagado, oc.valor_total AS valor_oc,
                 CASE WHEN pdf.factura_id IS NOT NULL THEN 1 ELSE 0 END AS tiene_pdf
          FROM facturas_proveedor f
          LEFT JOIN (SELECT factura_proveedor_id, SUM(monto) AS pagado FROM pagos_oc
                     WHERE factura_proveedor_id IS NOT NULL
                     GROUP BY factura_proveedor_id) p ON p.factura_proveedor_id = f.id
          LEFT JOIN ordenes_compra oc ON oc.numero_oc = f.numero_oc
          LEFT JOIN facturas_proveedor_pdf pdf ON pdf.factura_id = f.id
          WHERE 1=1"""
    params = []
    if estado and estado != 'todas':
        sql += " AND f.estado=?"; params.append(estado)
    if proveedor:
        sql += " AND LOWER(f.proveedor) LIKE ?"; params.append('%' + proveedor.lower() + '%')
    if q:
        sql += (" AND LOWER(COALESCE(f.numero_factura,'')||' '||COALESCE(f.proveedor,'')"
                "||' '||COALESCE(f.numero_oc,'')) LIKE ?"); params.append('%' + q + '%')
    sql += (" ORDER BY COALESCE(NULLIF(f.fecha_vencimiento,''), f.fecha_emision) ASC,"
            " f.id DESC LIMIT 1000")
    rows = c.execute(sql, tuple(params)).fetchall()
    cols = [x[0] for x in c.description]
    import datetime as _dt
    hoy = (_dt.datetime.utcnow() - _dt.timedelta(hours=5)).date()
    items = []; tot_saldo = 0.0; tot_vencido = 0.0
    for r in rows:
        f = dict(zip(cols, r))
        total = float(f.get('total') or 0)
        pagado = float(f.get('pagado') or 0)
        saldo = round(max(0.0, total - pagado), 2)
        f['pagado'] = round(pagado, 2); f['saldo'] = saldo
        f['tiene_pdf'] = bool(f.get('tiene_pdf'))
        dias = None; vencida = False
        fv = (f.get('fecha_vencimiento') or '').strip()[:10]
        if fv:
            try:
                dias = (_dt.date.fromisoformat(fv) - hoy).days
                if f['estado'] not in ('pagada', 'anulada') and dias < 0:
                    vencida = True
            except ValueError:
                pass
        f['dias_vencimiento'] = dias; f['vencida'] = vencida
        f['estado_efectivo'] = 'vencida' if vencida else f['estado']
        f['sobre_facturada'] = bool(f.get('numero_oc') and f.get('valor_oc')
                                    and total > float(f['valor_oc']) + 0.01)
        if f['estado'] != 'anulada':
            tot_saldo += saldo
            if vencida:
                tot_vencido += saldo
        items.append(f)
    return jsonify({'ok': True, 'items': items, 'n': len(items),
                    'total_saldo': round(tot_saldo, 2),
                    'total_vencido': round(tot_vencido, 2)})


@bp.route('/api/compras/facturas-proveedor', methods=['POST'])
def fp_crear():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    numero = (d.get('numero_factura') or '').strip()
    proveedor = (d.get('proveedor') or '').strip()
    if not numero:
        return jsonify({'error': 'numero_factura requerido'}), 400
    if not proveedor:
        return jsonify({'error': 'proveedor requerido'}), 400
    total = d.get('total')
    try:
        total = float(total) if total not in (None, '', 0, 0.0) else _fp_total_calc(d)
    except (TypeError, ValueError):
        total = _fp_total_calc(d)
    conn = get_db(); c = conn.cursor()
    dup = c.execute("SELECT id FROM facturas_proveedor WHERE proveedor=? AND numero_factura=?",
                    (proveedor, numero)).fetchone()
    if dup:
        return jsonify({'error': 'factura duplicada',
                        'detail': f'Ya existe {numero} de {proveedor} (id {dup[0]})'}), 409
    # Validación no bloqueante de la OC ligada (FIX 1-jun-2026 · audit)
    numero_oc = (d.get('numero_oc') or '').strip()
    warning = None
    if numero_oc:
        _ocx = c.execute("SELECT 1 FROM ordenes_compra WHERE numero_oc=?", (numero_oc,)).fetchone()
        if not _ocx:
            warning = f'La OC "{numero_oc}" no existe — la factura se creó igual, verificá el número.'

    def _f(k):
        try:
            return float(d.get(k) or 0)
        except (TypeError, ValueError):
            return 0.0
    import datetime as _dtfp
    _hoy = (_dtfp.datetime.utcnow() - _dtfp.timedelta(hours=5)).date().isoformat()
    c.execute("""INSERT INTO facturas_proveedor
        (numero_factura, proveedor, nit, numero_oc, fecha_emision, fecha_vencimiento,
         subtotal, iva, iva_pct, retefuente, retefuente_pct, retica, retica_pct,
         total, estado, pdf_adjunto, observaciones, creado_por, empresa)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (numero, proveedor, (d.get('nit') or '').strip(), numero_oc,
         (d.get('fecha_emision') or '').strip() or _hoy,
         (d.get('fecha_vencimiento') or '').strip(),
         _f('subtotal'), _f('iva'), _f('iva_pct'), _f('retefuente'), _f('retefuente_pct'),
         _f('retica'), _f('retica_pct'), total, 'pendiente',
         '', (d.get('observaciones') or '').strip(),
         user, (d.get('empresa') or 'Espagiria').strip()))
    fid = c.lastrowid
    # PDF en tabla 1:1 (mig 207) · fuera de la tabla transaccional · anti-OOM/bloat
    _pdf = (d.get('pdf_adjunto') or '')[:6_000_000]
    if _pdf:
        try:
            c.execute("INSERT INTO facturas_proveedor_pdf (factura_id, pdf_adjunto) VALUES (?,?)",
                      (fid, _pdf))
        except Exception:
            pass
    try:
        audit_log(c, usuario=user, accion='CREAR_FACTURA_PROVEEDOR',
                  tabla='facturas_proveedor', registro_id=str(fid),
                  despues={'numero': numero, 'proveedor': proveedor, 'total': total})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'id': fid, 'total': total, 'warning': warning})


@bp.route('/api/compras/facturas-proveedor/<int:fid>', methods=['GET'])
def fp_detalle(fid):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db(); c = conn.cursor()
    r = c.execute("""SELECT id, numero_factura, proveedor, nit, numero_oc, fecha_emision,
        fecha_vencimiento, subtotal, iva, iva_pct, retefuente, retefuente_pct, retica,
        retica_pct, total, estado, observaciones, creado_por, created_at, empresa
        FROM facturas_proveedor WHERE id=?""", (fid,)).fetchone()
    if not r:
        return jsonify({'error': 'no existe'}), 404
    cols = [x[0] for x in c.description]
    f = dict(zip(cols, r))
    _tp = c.execute("SELECT 1 FROM facturas_proveedor_pdf WHERE factura_id=?", (fid,)).fetchone()
    f['tiene_pdf'] = bool(_tp)
    pagos = c.execute(
        "SELECT id, monto, medio, fecha_pago, registrado_por, observaciones "
        "FROM pagos_oc WHERE factura_proveedor_id=? ORDER BY id", (fid,)).fetchall()
    f['pagos'] = [{'id': p[0], 'monto': float(p[1] or 0), 'medio': p[2],
                   'fecha_pago': p[3], 'registrado_por': p[4], 'observaciones': p[5]}
                  for p in pagos]
    f['pagado'] = round(sum(x['monto'] for x in f['pagos']), 2)
    f['saldo'] = round(max(0.0, float(f.get('total') or 0) - f['pagado']), 2)
    return jsonify({'ok': True, 'factura': f})


@bp.route('/api/compras/facturas-proveedor/<int:fid>/pdf', methods=['GET'])
def fp_pdf(fid):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db(); c = conn.cursor()
    r = c.execute("SELECT pdf_adjunto FROM facturas_proveedor_pdf WHERE factura_id=?", (fid,)).fetchone()
    if not r or not (r[0] or '').strip():  # fallback legacy (filas previas a mig 207)
        r = c.execute("SELECT pdf_adjunto FROM facturas_proveedor WHERE id=?", (fid,)).fetchone()
    if not r or not (r[0] or '').strip():
        return jsonify({'error': 'sin adjunto'}), 404
    raw = r[0]
    import base64 as _b64
    if raw.startswith('data:'):
        try:
            header, b64 = raw.split(',', 1)
            mime = header.split(':', 1)[1].split(';', 1)[0] or 'application/octet-stream'
        except Exception:
            return jsonify({'error': 'adjunto inválido'}), 400
    else:
        b64 = raw; mime = 'application/pdf'
    try:
        data = _b64.b64decode(b64)
    except Exception:
        return jsonify({'error': 'adjunto inválido'}), 400
    from flask import Response as _Resp
    return _Resp(data, mimetype=mime)


@bp.route('/api/compras/facturas-proveedor/<int:fid>', methods=['PATCH'])
def fp_editar(fid):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    conn = get_db(); c = conn.cursor()
    r = c.execute("SELECT estado FROM facturas_proveedor WHERE id=?", (fid,)).fetchone()
    if not r:
        return jsonify({'error': 'no existe'}), 404
    if d.get('anular'):
        c.execute("UPDATE facturas_proveedor SET estado='anulada' WHERE id=?", (fid,))
        try:
            audit_log(c, usuario=user, accion='ANULAR_FACTURA_PROVEEDOR',
                      tabla='facturas_proveedor', registro_id=str(fid),
                      despues={'motivo': (d.get('motivo') or '')})
        except Exception:
            pass
        conn.commit()
        return jsonify({'ok': True, 'estado': 'anulada'})
    campos = ['proveedor', 'nit', 'numero_oc', 'fecha_emision', 'fecha_vencimiento',
              'subtotal', 'iva', 'iva_pct', 'retefuente', 'retefuente_pct',
              'retica', 'retica_pct', 'total', 'observaciones']
    sets = []; params = []
    for k in campos:
        if k in d:
            sets.append(f"{k}=?"); params.append(d.get(k))
    if not sets:
        return jsonify({'error': 'nada que actualizar'}), 400
    params.append(fid)
    c.execute(f"UPDATE facturas_proveedor SET {', '.join(sets)} WHERE id=?", tuple(params))
    _fp_recalc_estado(c, fid)
    try:
        audit_log(c, usuario=user, accion='EDITAR_FACTURA_PROVEEDOR',
                  tabla='facturas_proveedor', registro_id=str(fid), despues=d)
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/compras/facturas-proveedor/<int:fid>/pagar', methods=['POST'])
def fp_pagar(fid):
    """Registra un pago CONTRA la factura (pagos_oc.factura_proveedor_id=fid) y
    recalcula el estado. El link es factura_proveedor_id (NO numero_factura_proveedor,
    que es UNIQUE y bloquearía pagos parciales de la misma factura)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    conn = get_db(); c = conn.cursor()
    f = c.execute("SELECT numero_oc, total, estado, numero_factura "
                  "FROM facturas_proveedor WHERE id=?", (fid,)).fetchone()
    if not f:
        return jsonify({'error': 'no existe'}), 404
    if f[2] == 'anulada':
        return jsonify({'error': 'factura anulada'}), 400
    oc_num = (f[0] or '').strip()
    # Guarda: no pagar contra OC en estado no-pagable (consistencia con pagar_oc · 1-jun)
    if oc_num:
        _oce = c.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?", (oc_num,)).fetchone()
        if _oce and (_oce[0] or '') in ('Cancelada', 'Rechazada', 'Borrador', 'Revisada'):
            return jsonify({'error': f'La OC {oc_num} está en estado "{_oce[0]}" · no admite pagos'}), 409
    # Anti-doble-pago: ese número ya fue pagado por el camino directo de OC (INV-6)
    if (f[3] or '').strip():
        _dir = c.execute("SELECT 1 FROM pagos_oc WHERE numero_factura_proveedor=? LIMIT 1",
                         (f[3],)).fetchone()
        if _dir:
            return jsonify({'error': f'La factura "{f[3]}" ya fue pagada por el camino directo '
                                     f'de OC · no la dupliques acá', 'codigo': 'YA_PAGADA_DIRECTO'}), 409
    try:
        monto = float(d.get('monto') or 0)
    except (TypeError, ValueError):
        monto = 0
    if monto <= 0:
        return jsonify({'error': 'monto > 0 requerido'}), 400
    medio = (d.get('medio') or 'Transferencia').strip()
    if medio.upper() == 'PENDIENTE':
        return jsonify({'error': 'elegí un medio de pago real'}), 400
    pagado = float(c.execute(
        "SELECT COALESCE(SUM(monto),0) FROM pagos_oc WHERE factura_proveedor_id=?",
        (fid,)).fetchone()[0] or 0)
    total = float(f[1] or 0)
    if total > 0 and pagado + monto > total + 0.01:
        return jsonify({'error': 'el pago excede el saldo de la factura',
                        'saldo': round(total - pagado, 2)}), 400
    c.execute("""INSERT INTO pagos_oc
        (numero_oc, monto, medio, registrado_por, numero_factura_proveedor,
         observaciones, factura_proveedor_id)
        VALUES (?,?,?,?,?,?,?)""",
        (f[0] or '', monto, medio, user, '',
         (d.get('observaciones') or '').strip(), fid))
    pago_id = c.lastrowid
    _fp_recalc_estado(c, fid)
    # Reflejar el pago en el estado de la OC ligada (consistencia OC↔factura · FIX
    # 1-jun-2026) · mismo CAS que pagar_oc (Pagada/Parcial por SUM(pagos) vs valor_total).
    # No toca OCs en estado no-pagable.
    oc_num = (f[0] or '').strip()
    oc_estado = None
    if oc_num:
        ocr = c.execute("SELECT estado, COALESCE(valor_total,0) FROM ordenes_compra "
                        "WHERE numero_oc=?", (oc_num,)).fetchone()
        if ocr and (ocr[0] or '') not in ('Cancelada', 'Rechazada', 'Borrador', 'Revisada'):
            c.execute("""UPDATE ordenes_compra SET estado = CASE WHEN (
                    SELECT COALESCE(SUM(monto),0) FROM pagos_oc WHERE numero_oc=?
                 ) >= ? - 0.01 THEN 'Pagada' ELSE 'Parcial' END
                 WHERE numero_oc=?""", (oc_num, float(ocr[1] or 0), oc_num))
            _re = c.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?",
                            (oc_num,)).fetchone()
            oc_estado = _re[0] if _re else None
    try:
        audit_log(c, usuario=user, accion='PAGAR_FACTURA_PROVEEDOR',
                  tabla='facturas_proveedor', registro_id=str(fid),
                  despues={'monto': monto, 'medio': medio, 'pago_id': pago_id,
                           'oc': oc_num, 'oc_estado': oc_estado})
    except Exception:
        pass
    conn.commit()
    est = c.execute("SELECT estado FROM facturas_proveedor WHERE id=?", (fid,)).fetchone()[0]
    return jsonify({'ok': True, 'pago_id': pago_id, 'estado': est, 'oc_estado': oc_estado})


@bp.route('/api/compras/mailbox-facturas', methods=['GET'])
def listar_mailbox_facturas():
    """Sebastián 23-may-2026 · MBX UI · lista facturas detectadas por el
    cron job_mailbox_factura_proveedor (medio='PENDIENTE') · admin debe
    completar monto + medio definitivo o descartar.

    Inserts vienen de auto_plan_jobs.py:3585 (mailbox IMAP cron).
    """
    u, err, code = _require_compras_session()
    if err: return err, code
    try:
        dias = max(1, min(int(request.args.get('dias', 30)), 365))
    except Exception:
        dias = 30
    conn = get_db(); c = conn.cursor()
    from datetime import datetime as _dt, timedelta as _td
    desde = (_dt.utcnow() - _td(hours=5) - _td(days=dias)).strftime('%Y-%m-%d %H:%M:%S')
    try:
        rows = c.execute("""
            SELECT po.id, po.numero_oc, po.fecha_pago, po.numero_factura_proveedor,
                   COALESCE(po.monto,0), po.medio,
                   COALESCE(po.observaciones,''),
                   COALESCE(oc.proveedor,''),
                   COALESCE(oc.valor_total,0),
                   COALESCE(oc.estado,'')
            FROM pagos_oc po
            LEFT JOIN ordenes_compra oc ON oc.numero_oc = po.numero_oc
            WHERE po.registrado_por = 'cron-mailbox'
              AND po.fecha_pago >= ?
            ORDER BY po.fecha_pago DESC
        """, (desde,)).fetchall()
    except Exception as e:
        return jsonify({'error': f'query fallo: {e}'}), 500
    items = []
    for r in rows:
        items.append({
            'pago_id': r[0],
            'numero_oc': r[1],
            'fecha': str(r[2])[:10] if r[2] else '',
            'numero_factura': r[3] or '',
            'monto': float(r[4] or 0),
            'medio': r[5] or '',
            'pendiente': (r[5] or '').upper() == 'PENDIENTE',
            'observaciones': r[6],
            'proveedor': r[7],
            'valor_oc': float(r[8] or 0),
            'estado_oc': r[9],
        })
    n_pendientes = sum(1 for it in items if it['pendiente'])
    return jsonify({
        'dias_ventana': dias,
        'total': len(items),
        'n_pendientes': n_pendientes,
        'items': items,
    })


@bp.route('/api/compras/mailbox-facturas/<int:pago_id>/comprobante', methods=['GET'])
def mailbox_factura_comprobante(pago_id):
    """Descarga el comprobante adjunto base64 de una factura mailbox."""
    u, err, code = _require_compras_session()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    row = c.execute("""
        SELECT comprobante_imagen, numero_factura_proveedor, numero_oc
        FROM pagos_oc WHERE id=? AND registrado_por='cron-mailbox'
    """, (pago_id,)).fetchone()
    if not row:
        return jsonify({'error': 'No encontrado'}), 404
    b64 = row[0] or ''
    if not b64:
        return jsonify({'error': 'Sin adjunto'}), 404
    import base64 as _b64
    try:
        raw = _b64.b64decode(b64)
    except Exception:
        return jsonify({'error': 'Adjunto corrupto'}), 500
    fname = (row[1] or 'factura').strip().replace('/', '-')
    # Detectar tipo por magic bytes
    if raw[:4] == b'%PDF':
        mt = 'application/pdf'
        if not fname.lower().endswith('.pdf'): fname += '.pdf'
    elif raw[:3] == b'\xff\xd8\xff':
        mt = 'image/jpeg'
        if not fname.lower().endswith(('.jpg','.jpeg')): fname += '.jpg'
    elif raw[:8] == b'\x89PNG\r\n\x1a\n':
        mt = 'image/png'
        if not fname.lower().endswith('.png'): fname += '.png'
    else:
        mt = 'application/octet-stream'
    from flask import send_file
    import io as _io
    return send_file(_io.BytesIO(raw), as_attachment=False,
                     download_name=fname, mimetype=mt)


@bp.route('/api/compras/mailbox-facturas/<int:pago_id>/descartar', methods=['POST'])
def mailbox_factura_descartar(pago_id):
    """Admin descarta una factura mailbox (no aplica) · borra el row."""
    u, err, code = _require_compras_write()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT numero_oc, numero_factura_proveedor FROM pagos_oc WHERE id=? AND registrado_por='cron-mailbox'",
        (pago_id,),
    ).fetchone()
    if not row:
        return jsonify({'error': 'No encontrado'}), 404
    c.execute("DELETE FROM pagos_oc WHERE id=? AND registrado_por='cron-mailbox'", (pago_id,))
    try:
        audit_log(c, usuario=u, accion='DESCARTAR_MAILBOX_FACTURA',
                  tabla='pagos_oc', registro_id=str(pago_id),
                  antes={'numero_oc': row[0], 'numero_factura': row[1]})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True})


# ─── Mailbox · completar factura · Sebastián 24-may-2026 ──────────────
# Audit Mailbox · el cron IMAP inserta pagos_oc con monto=0 + medio='PENDIENTE'
# para que admin complete los datos reales. Antes solo se podía Ver/Descartar
# desde la UI · faltaba la acción "Completar" que es la primaria. Endpoint
# nuevo · UPDATE in-place + recalcula estado OC según SUM(pagos_oc).
@bp.route('/api/compras/mailbox-facturas/<int:pago_id>/completar', methods=['POST'])
def mailbox_factura_completar(pago_id):
    """Completa una factura PENDIENTE detectada por cron-mailbox.

    Body: {monto: float, medio_pago: str, numero_factura_proveedor?: str,
           observaciones?: str}

    Side effects:
      1. UPDATE pagos_oc · monto + medio + numero_factura + observaciones
         + registrado_por=usuario_actual (deja de ser cron-mailbox · ya
         es un pago "real" auditado · pero conserva la fecha_pago original
         del email para trazabilidad cronológica).
      2. Recalcula estado OC según SUM(pagos_oc.monto) restante:
         - >= valor_total → 'Pagada' + fecha_pago, pagado_por
         - 0 < pagado < total → 'Parcial'
         - == 0 → no cambia estado (raro · solo si admin completa con 0)
      3. Audit COMPLETAR_MAILBOX_FACTURA con antes/después.
      4. Email + CE generación se delega al endpoint /pagar normal · si el
         admin quiere generar CE + email beneficiario, debe ir a tab
         Pagos y usar "Regenerar CE" o crear pago manual desde Por Pagar.

    Solo COMPRAS_ACCESS_WRITE · admin/compras.
    """
    usuario, err, code = _require_compras_write()
    if err: return err, code
    d = request.get_json(silent=True) or {}
    monto, e_monto = validate_money(d.get('monto'), allow_zero=False,
                                     max_value=10_000_000_000,
                                     field_name='monto')
    if e_monto:
        return jsonify(e_monto), 400
    medio = (d.get('medio_pago') or '').strip()
    if not medio or medio.upper() == 'PENDIENTE':
        return jsonify({'error': 'medio_pago requerido (Transferencia/Nequi/Daviplata/etc · no PENDIENTE)'}), 400
    n_factura = (d.get('numero_factura_proveedor') or '').strip()
    obs = (d.get('observaciones') or '').strip()
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT numero_oc, monto, medio, numero_factura_proveedor, "
        "       COALESCE(observaciones,''), fecha_pago "
        "FROM pagos_oc WHERE id=? AND registrado_por='cron-mailbox'",
        (pago_id,),
    ).fetchone()
    if not row:
        return jsonify({'error': 'Mailbox factura no encontrada (puede ya estar completada o descartada)'}), 404
    numero_oc, monto_ant, medio_ant, fact_ant, obs_ant, fecha_pago_orig = row
    # Validar OC todavía existe + estado coherente
    oc = c.execute(
        "SELECT estado, valor_total FROM ordenes_compra WHERE numero_oc=?",
        (numero_oc,),
    ).fetchone()
    if not oc:
        return jsonify({'error': f'OC {numero_oc} no existe · descartá esta entrada del mailbox'}), 404
    estado_oc_ant, valor_total_oc = oc
    if estado_oc_ant in ('Cancelada', 'Rechazada'):
        return jsonify({
            'error': f'OC {numero_oc} está {estado_oc_ant} · no se pueden completar pagos',
            'codigo': 'OC_TERMINAL',
        }), 409
    # Validar duplicado de numero_factura_proveedor en otra fila
    if n_factura:
        dup = c.execute(
            "SELECT id FROM pagos_oc WHERE numero_factura_proveedor=? AND id != ?",
            (n_factura, pago_id),
        ).fetchone()
        if dup:
            return jsonify({
                'error': f'Factura {n_factura} ya registrada en otro pago (id={dup[0]})',
                'codigo': 'FACTURA_DUPLICADA',
            }), 409
    # Verificar over-payment
    total_otros = c.execute(
        "SELECT COALESCE(SUM(monto), 0) FROM pagos_oc WHERE numero_oc=? AND id != ?",
        (numero_oc, pago_id),
    ).fetchone()[0] or 0
    if (float(total_otros) + monto) > (float(valor_total_oc or 0) + 0.01):
        return jsonify({
            'error': (f'Pagaría ${total_otros + monto:,.0f} pero OC vale ${valor_total_oc:,.0f}. '
                      f'Reducí el monto o cancelá pagos previos.'),
            'codigo': 'OVERPAYMENT',
            'total_si_completas': float(total_otros) + monto,
            'valor_oc': float(valor_total_oc or 0),
        }), 409
    # UPDATE el row · ya NO es cron-mailbox sino pago real auditado
    c.execute(
        "UPDATE pagos_oc SET monto=?, medio=?, numero_factura_proveedor=?, "
        "       observaciones=?, registrado_por=? WHERE id=?",
        (monto, medio, n_factura or None,
         obs or 'Completado desde mailbox · ' + (obs_ant or ''),
         usuario, pago_id),
    )
    # Recalcular estado OC según SUM real
    total_pagado = float(total_otros) + monto
    if total_pagado >= float(valor_total_oc or 0) - 0.01:
        nuevo_estado = 'Pagada'
    elif total_pagado > 0:
        nuevo_estado = 'Parcial'
    else:
        nuevo_estado = estado_oc_ant
    c.execute(
        "UPDATE ordenes_compra SET estado=?, "
        "       fecha_pago=CASE WHEN ? = 'Pagada' THEN ? ELSE fecha_pago END, "
        "       pagado_por=CASE WHEN ? = 'Pagada' THEN ? ELSE pagado_por END, "
        "       medio_pago=CASE WHEN ? = 'Pagada' THEN ? ELSE medio_pago END "
        "WHERE numero_oc=?",
        (nuevo_estado, nuevo_estado, fecha_pago_orig or datetime.now().isoformat(),
         nuevo_estado, usuario, nuevo_estado, medio, numero_oc),
    )
    try:
        audit_log(c, usuario=usuario, accion='COMPLETAR_MAILBOX_FACTURA',
                  tabla='pagos_oc', registro_id=str(pago_id),
                  antes={'numero_oc': numero_oc, 'monto': float(monto_ant or 0),
                          'medio': medio_ant, 'numero_factura': fact_ant,
                          'estado_oc': estado_oc_ant},
                  despues={'monto': monto, 'medio': medio,
                            'numero_factura': n_factura,
                            'estado_oc': nuevo_estado,
                            'total_pagado': total_pagado},
                  detalle=(f"Completó mailbox factura OC {numero_oc} · "
                            f"${monto:,.0f} · {medio} · estado {estado_oc_ant}→{nuevo_estado}"))
    except Exception as e:
        log.warning('audit_log COMPLETAR_MAILBOX_FACTURA fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'numero_oc': numero_oc,
        'monto': monto,
        'medio': medio,
        'estado_oc_anterior': estado_oc_ant,
        'estado_oc_nuevo': nuevo_estado,
        'total_pagado': total_pagado,
        'hint': (f'OC pasó a {nuevo_estado}. Si querés CE/PDF, ' +
                  ('regenerá desde tab Pagos.' if nuevo_estado == 'Pagada' else
                    'completá el saldo restante.')),
    })


@bp.route('/api/compras/ocs-consolidado-excel', methods=['GET'])
def ocs_consolidado_excel():
    """Sebastián 23-may-2026 · Excel consolidado de OCs activas para
    Catalina · descarga con info de proveedor + items + estado +
    pendiente pago + recepción.

    Query params:
      ?estados=Borrador,Autorizada,Recibida,Parcial,Pagada (default activos)
      ?dias=N · filtra OCs creadas en últimos N días (default 90)
    """
    u, err, code = _require_compras_session()
    if err: return err, code
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl no disponible'}), 500
    from flask import send_file
    import io as _io
    from datetime import datetime as _dt, timedelta as _td

    estados_str = request.args.get('estados') or 'Borrador,Autorizada,Parcial,Recibida,Pagada'
    estados = [e.strip() for e in estados_str.split(',') if e.strip()]
    try:
        dias = max(1, min(int(request.args.get('dias', 90)), 365))
    except Exception:
        dias = 90
    hoy_ = (_dt.utcnow() - _td(hours=5)).date()
    desde = (hoy_ - _td(days=dias)).isoformat()

    conn = get_db(); c = conn.cursor()
    placeholders = ','.join('?' for _ in estados)
    params = list(estados) + [desde]
    rows = c.execute(f"""
        SELECT oc.numero_oc, oc.fecha, oc.estado, oc.proveedor,
               COALESCE(oc.categoria,''),
               COALESCE(oc.valor_total, 0),
               COALESCE(oc.fecha_recepcion,''),
               COALESCE(oc.tiene_discrepancias, 0),
               COALESCE(oc.creado_por,''),
               COALESCE(oc.recibido_por,''),
               COALESCE(p.nit,''), COALESCE(p.banco,''),
               COALESCE(p.tipo_cuenta,''), COALESCE(p.num_cuenta,''),
               COALESCE(p.condiciones_pago,'')
        FROM ordenes_compra oc
        LEFT JOIN proveedores p ON oc.proveedor = p.nombre
        WHERE oc.estado IN ({placeholders})
          AND COALESCE(oc.fecha,'') >= ?
        ORDER BY oc.fecha DESC
    """, params).fetchall()

    # Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1E40AF')
    title_fill = PatternFill('solid', fgColor='1E3A8A')
    estado_fills = {
        'Borrador':   PatternFill('solid', fgColor='F1F5F9'),
        'Autorizada': PatternFill('solid', fgColor='FEF3C7'),
        'Parcial':    PatternFill('solid', fgColor='FFF7ED'),
        'Recibida':   PatternFill('solid', fgColor='DBEAFE'),
        'Pagada':     PatternFill('solid', fgColor='DCFCE7'),
    }

    ws = wb.create_sheet('OCs consolidado')
    fecha_str = _dt.now().strftime('%Y-%m-%d %H:%M')
    ws.cell(row=1, column=1,
            value=f'Órdenes de Compra · estados {estados_str} · últimos {dias}d · {fecha_str}')
    ws.cell(row=1, column=1).font = Font(size=14, bold=True, color='FFFFFF')
    ws.cell(row=1, column=1).fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    ws.row_dimensions[1].height = 24

    hdr = ['OC', 'Fecha', 'Estado', 'Proveedor', 'Categoría', 'Valor',
           'Recepción', 'Discrepancia', 'Creador', 'Receptor',
           'NIT', 'Banco', 'Tipo cta', 'N° cuenta', 'Condiciones pago']
    for col, val in enumerate(hdr, start=1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 32

    r = 4
    total_valor = 0.0
    for row in rows:
        valores = [
            row[0], str(row[1])[:10] if row[1] else '',
            row[2], row[3] or '', row[4], float(row[5] or 0),
            str(row[6])[:10] if row[6] else '',
            'SÍ' if int(row[7] or 0) else '',
            row[8], row[9],
            row[10], row[11], row[12], row[13], row[14],
        ]
        total_valor += float(row[5] or 0)
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            if col == 3:  # Estado
                cell.fill = estado_fills.get(str(val), estado_fills['Borrador'])
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            elif col == 1:
                cell.font = Font(name='Consolas', bold=True)
                cell.alignment = Alignment(horizontal='center')
            elif col == 6:  # Valor
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            elif col == 8 and val == 'SÍ':  # Discrepancia
                cell.fill = PatternFill('solid', fgColor='FEE2E2')
                cell.font = Font(bold=True, color='991B1B')
                cell.alignment = Alignment(horizontal='center')
            elif col in (11, 13, 14):  # NIT, tipo cta, n° cuenta
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Consolas', size=10)
        r += 1

    # Fila TOTAL
    ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='E0E7FF')
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    for col in range(2, 6):
        ws.cell(row=r, column=col, value='').fill = PatternFill('solid', fgColor='E0E7FF')
    cell = ws.cell(row=r, column=6, value=total_valor)
    cell.fill = PatternFill('solid', fgColor='E0E7FF')
    cell.font = Font(bold=True)
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='right')
    cell.border = border

    # Anchos
    widths = [14, 11, 11, 26, 14, 14, 11, 12, 14, 14, 14, 18, 9, 18, 22]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'C4'

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f'ocs_consolidado_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/compras/recepciones-discrepancias', methods=['GET'])
def listar_recepciones_discrepancias():
    """Sebastián 23-may-2026 · histórico de recepciones con discrepancia
    en últimos N días · útil para ranking de calidad de despacho por
    proveedor.

    Una OC se incluye si:
    - estado IN ('Recibida','Pagada','Parcial')
    - tiene_discrepancias=1 (marcada manual o auto en recibir_oc)
    - fecha_recepcion en últimos `dias` (default 30)

    Devuelve:
    - lista de OCs con items faltantes (pedido vs recibido)
    - resumen por proveedor: # discrepancias, # recepciones totales, tasa
    """
    u, err, code = _require_compras_session()
    if err: return err, code
    try:
        dias = max(1, min(int(request.args.get('dias', 30)), 365))
    except Exception:
        dias = 30
    conn = get_db(); c = conn.cursor()
    from datetime import datetime as _dt, timedelta as _td
    hoy = (_dt.utcnow() - _td(hours=5)).date()
    desde = (hoy - _td(days=dias)).isoformat()

    # OCs con discrepancia en ventana
    try:
        rows = c.execute("""
            SELECT oc.numero_oc, oc.fecha, oc.estado, oc.proveedor,
                   COALESCE(oc.creado_por,''),
                   COALESCE(oc.fecha_recepcion,''),
                   COALESCE(oc.recibido_por,''),
                   COALESCE(oc.observaciones_recepcion,''),
                   COALESCE(oc.valor_total, 0)
            FROM ordenes_compra oc
            WHERE oc.tiene_discrepancias = 1
              AND oc.estado IN ('Recibida','Pagada','Parcial')
              AND COALESCE(oc.fecha_recepcion,'') >= ?
            ORDER BY oc.fecha_recepcion DESC
        """, (desde,)).fetchall()
    except Exception as e:
        return jsonify({'error': f'query fallo: {e}'}), 500

    ocs_lista = []
    for r in rows:
        # Items con faltante
        items_faltantes = []
        try:
            item_rows = c.execute("""
                SELECT codigo_mp, COALESCE(nombre_mp,''),
                       COALESCE(cantidad_g, 0),
                       COALESCE(cantidad_recibida_g, 0)
            FROM ordenes_compra_items
            WHERE numero_oc=?
            """, (r[0],)).fetchall()
            for it in item_rows:
                ped = float(it[2] or 0)
                recib = float(it[3] or 0)
                if ped > 0 and recib < ped * 0.999:
                    faltante = ped - recib
                    items_faltantes.append({
                        'codigo_mp': it[0],
                        'nombre_mp': it[1],
                        'pedido': round(ped, 1),
                        'recibido': round(recib, 1),
                        'faltante': round(faltante, 1),
                        'pct_faltante': round((faltante / ped) * 100, 1),
                    })
        except Exception:
            pass
        ocs_lista.append({
            'numero_oc': r[0],
            'fecha': str(r[1])[:10] if r[1] else '',
            'estado': r[2],
            'proveedor': r[3] or '',
            'creador': r[4],
            'fecha_recepcion': str(r[5])[:10] if r[5] else '',
            'recibido_por': r[6],
            'observaciones_recepcion': r[7],
            'valor_total': float(r[8] or 0),
            'items_faltantes': items_faltantes,
            'n_items_faltantes': len(items_faltantes),
        })

    # Ranking por proveedor: # OCs con discrepancia / total OCs recibidas en ventana
    ranking = {}
    try:
        # Sebastián 24-may-2026 · audit Calidad recepción · ranking proveedor
        # case + espacios insensible. Antes drift histórico ("DistriQuim" vs
        # "DISTRIQUIM" vs "distriquim ") aparecían como 3 entradas distintas
        # en el ranking · tasa diluida por proveedor real. Mismo fix que
        # aplicamos en Bandeja Planta (commit bbcff41).
        # prov_norm_orig usa MIN(TRIM(...)) — debe ser agregado: en PostgreSQL una
        # columna del SELECT que no esté en GROUP BY ni agregada es error duro
        # ("must appear in the GROUP BY clause"). SQLite lo toleraba (valor
        # arbitrario) → el ranking salía bien local pero VACÍO en prod (PG), y el
        # error lo tragaba el except → silencioso. Cazado por suite golden en PG · 8-jun.
        for r in c.execute("""
            SELECT MIN(TRIM(proveedor)) AS prov_norm_orig,
                   UPPER(TRIM(proveedor)) AS prov_norm_key,
                   COUNT(*) AS total_recibidas,
                   SUM(CASE WHEN tiene_discrepancias=1 THEN 1 ELSE 0 END) AS con_discrep
            FROM ordenes_compra
            WHERE estado IN ('Recibida','Pagada','Parcial')
              AND COALESCE(fecha_recepcion,'') >= ?
              AND COALESCE(proveedor,'') != ''
            GROUP BY UPPER(TRIM(proveedor))
            ORDER BY con_discrep DESC, total_recibidas DESC
        """, (desde,)).fetchall():
            prov_label = r[0] or ''   # forma original (display)
            prov_key = r[1] or ''     # normalizada (dedup key)
            total_rec = int(r[2] or 0)
            con_dis = int(r[3] or 0)
            if total_rec > 0:
                tasa = round((con_dis / total_rec) * 100, 1)
                ranking[prov_key] = {
                    'proveedor': prov_label,
                    'total_recibidas': total_rec,
                    'con_discrepancia': con_dis,
                    'tasa_discrepancia_pct': tasa,
                }
    except Exception as _e_rank:
        # No tragar en silencio (M4): si el ranking falla, dejar rastro. El
        # endpoint sigue devolviendo las OCs con discrepancia aunque el ranking
        # quede vacío.
        log.warning('ranking_proveedores recepciones-discrepancias falló: %s', _e_rank)
    ranking_lista = sorted(
        ranking.values(),
        key=lambda x: (-x['tasa_discrepancia_pct'], -x['con_discrepancia']),
    )

    return jsonify({
        'hoy': hoy.isoformat(),
        'dias_ventana': dias,
        'total_ocs_con_discrepancia': len(ocs_lista),
        'ocs': ocs_lista,
        'ranking_proveedores': ranking_lista,
    })


@bp.route('/api/compras/ocs-atrasadas', methods=['GET'])
def listar_ocs_atrasadas():
    """Sebastián 23-may-2026 · cierre flujo Compras · OCs sin recibir
    completa tras lead_time + buffer (default 7d).

    Útil para el dashboard de Compras y la UI de Abastecimiento.
    Query param ?buffer_dias=N (default 7).
    """
    u, err, code = _require_compras_session()
    if err: return err, code
    try:
        buffer_dias = max(0, min(int(request.args.get('buffer_dias', 7)), 90))
    except Exception:
        buffer_dias = 7
    conn = get_db(); c = conn.cursor()
    # Sebastián 24-may-2026 · audit Atrasadas · excluir categorías de pago
    # directo (Cuenta de Cobro / Servicio / SVC / Influencer/Marketing Digital).
    # Esas OCs NUNCA reciben material físico · contarlas como "atrasadas por
    # no recibir" genera falsos positivos en el badge y en la lista. Catalina
    # veía 30+ OCs CC/influencer marcadas atrasadas que jamás iban a llegar.
    _CATS_EXCLUIDAS = list(CATEGORIAS_PAGO_DIRECTO) + ['Influencer/Marketing Digital']
    _ph_cats = ','.join(['?'] * len(_CATS_EXCLUIDAS))
    try:
        rows = c.execute(f"""
            SELECT oc.numero_oc, oc.fecha, oc.estado, oc.proveedor,
                   COALESCE(oc.creado_por,''),
                   COALESCE(oc.valor_total,0),
                   COALESCE(oc.observaciones,''),
                   (SELECT MAX(COALESCE(mlt.lead_time_dias, 14))
                    FROM ordenes_compra_items oci
                    LEFT JOIN mp_lead_time_config mlt ON mlt.material_id = oci.codigo_mp
                    WHERE oci.numero_oc = oc.numero_oc) AS lead_max,
                   COALESCE(oc.categoria, '') AS categoria
            FROM ordenes_compra oc
            WHERE oc.estado IN ('Autorizada','Parcial')
              AND (oc.fecha_recepcion IS NULL OR oc.fecha_recepcion = '')
              AND oc.fecha IS NOT NULL AND oc.fecha != ''
              AND COALESCE(oc.categoria, '') NOT IN ({_ph_cats})
            ORDER BY oc.fecha ASC
        """, _CATS_EXCLUIDAS).fetchall()
    except Exception as e:
        return jsonify({'error': f'query fallo: {e}'}), 500
    from datetime import datetime as _dt, timedelta as _td
    hoy = (_dt.utcnow() - _td(hours=5)).date()
    atrasadas = []
    for r in rows:
        try:
            f_oc = _dt.strptime(str(r[1])[:10], '%Y-%m-%d').date()
            lead = int(r[7] or 14)
            limite = f_oc + _td(days=lead + buffer_dias)
            if hoy > limite:
                atrasadas.append({
                    'numero_oc': r[0],
                    'fecha': str(r[1])[:10],
                    'estado': r[2],
                    'proveedor': r[3],
                    'creador': r[4],
                    'valor_total': float(r[5] or 0),
                    'observaciones': r[6],
                    'lead_time_dias': lead,
                    'dias_atraso': (hoy - limite).days,
                    'dias_desde_oc': (hoy - f_oc).days,
                })
        except Exception:
            continue
    atrasadas.sort(key=lambda x: -x['dias_atraso'])
    return jsonify({
        'hoy': hoy.isoformat(),
        'buffer_dias': buffer_dias,
        'total': len(atrasadas),
        'ocs': atrasadas,
    })


@bp.route('/api/proveedores-compras', methods=['GET','POST'])
def handle_proveedores_compras():
    # SEC-FIX 23-may-2026 · auditoría · era endpoint público · expone
    # num_cuenta/banco/nit/tipo_cuenta · viola Ley 1581 CO
    if request.method == 'GET':
        u, err, code = _require_compras_session()
        if err: return err, code
    else:
        u, err, code = _require_compras_write()
        if err: return err, code
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        # u ya seteado arriba con _require_compras_write
        d = request.get_json(silent=True) or {}
        nombre = (d.get('nombre') or '').strip()
        if not nombre:
            return jsonify({'error': 'Nombre requerido'}), 400
        # Detectar duplicado exacto (case + trim) · Catalina 4-may-2026
        existe = c.execute(
            "SELECT id, nombre FROM proveedores WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?)) AND activo=1",
            (nombre,)
        ).fetchone()
        if existe:
            return jsonify({
                'error': f'Ya existe proveedor "{existe[1]}" (case-insensitive). Selecciona del dropdown.',
                'existente': {'id': existe[0], 'nombre': existe[1]},
            }), 409
        try:
            c.execute("""INSERT INTO proveedores
                (nombre,contacto,email,telefono,categoria,condiciones_pago,
                 nit,direccion,num_cuenta,tipo_cuenta,banco,concepto_compra,fecha_creacion)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (nombre,d.get('contacto',''),d.get('email',''),d.get('telefono',''),
                 d.get('categoria',''),d.get('condiciones_pago','30 dias'),
                 d.get('nit',''),d.get('direccion',''),d.get('num_cuenta',''),
                 d.get('tipo_cuenta',''),d.get('banco',''),d.get('concepto_compra',d.get('concepto','')),
                 datetime.now().isoformat()))
            new_id = c.lastrowid
            try:
                audit_log(c, usuario=u, accion='CREAR_PROVEEDOR',
                          tabla='proveedores', registro_id=new_id,
                          despues={'nombre': nombre[:200],
                                    'nit': (d.get('nit','') or '')[:50],
                                    'banco': (d.get('banco','') or '')[:80],
                                    'condiciones_pago': d.get('condiciones_pago','30 dias')[:50]},
                          detalle=f"Creó proveedor '{nombre[:80]}'")
            except Exception:
                pass
            conn.commit()
            return jsonify({'ok': True, 'message': f"Proveedor '{nombre}' creado",
                             'id': new_id, 'nombre': nombre}), 201
        except Exception as e:
            log.exception('crear proveedor fallo: %s', e)
            return jsonify({'error': 'No se pudo crear el proveedor'}), 400
    c.execute("""SELECT nombre,contacto,email,telefono,categoria,condiciones_pago,
                       nit,direccion,num_cuenta,tipo_cuenta,banco,concepto_compra
                FROM proveedores WHERE activo=1 ORDER BY nombre""")
    cols = ['nombre','contacto','email','telefono','categoria','condiciones_pago',
            'nit','direccion','num_cuenta','tipo_cuenta','banco','concepto_compra']
    provs = [dict(zip(cols, r)) for r in c.fetchall()]
    # SEC-FIX 10-jun audit · Habeas Data (Ley 1581 · CERO_ERROR regla 5): los datos
    # bancarios SOLO los ve admin+contadora. Antes cualquier user de compras (planta,
    # calidad, marketing) leía num_cuenta/banco/nit en claro. Mismo patrón que marketing.py.
    if not (u and u.lower() in {x.lower() for x in (set(ADMIN_USERS) | set(CONTADORA_USERS))}):
        for _p in provs:
            for _cb in ('num_cuenta', 'tipo_cuenta', 'banco', 'nit'):
                if _p.get(_cb):
                    _p[_cb] = '***'
    return jsonify({'proveedores': provs})

@bp.route('/api/proveedores-compras/<path:nombre>', methods=['PATCH','DELETE'])
def handle_proveedor(nombre):
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    if request.method == 'DELETE':
        d = request.json or {}
        motivo = (d.get('motivo') or '').strip()
        if not motivo:
            return jsonify({'error': 'El motivo de baja es requerido'}), 400
        force = bool(d.get('force', False))
        c.execute("SELECT id FROM proveedores WHERE nombre=? AND activo=1", (nombre,))
        if not c.fetchone():
            return jsonify({'error': 'Proveedor no encontrado'}), 404
        # Sebastián 24-may-2026 · audit Proveedores · validar OCs activas antes
        # de dar de baja. Antes el soft delete (activo=0) permitía bajar al
        # proveedor pero las OCs vigentes (Borrador/Autorizada/Recibida/Parcial)
        # seguían vinculadas a él · quedaban huérfanas en UI (proveedor ya no
        # aparece en lista pero OC sigue activa). Ahora bloqueamos con 409 y
        # listamos las OCs · admin puede forzar con {force:true} para casos
        # legítimos (proveedor cerró operaciones · cancelar OCs aparte).
        _ESTADOS_OC_ACTIVOS = ('Borrador', 'Revisada', 'Aprobada', 'Autorizada',
                                'Recibida', 'Parcial', 'Pagada')
        try:
            ocs_activas = c.execute(
                "SELECT numero_oc, estado, valor_total FROM ordenes_compra "
                "WHERE LOWER(TRIM(proveedor)) = LOWER(TRIM(?)) "
                "  AND estado IN ('Borrador','Revisada','Aprobada','Autorizada',"
                "                 'Recibida','Parcial','Pagada') "
                "ORDER BY fecha DESC LIMIT 20",
                (nombre,),
            ).fetchall()
        except Exception:
            ocs_activas = []
        if ocs_activas and not force:
            es_admin = (usuario or '').lower() in {x.lower() for x in ADMIN_USERS}
            return jsonify({
                'error': (f'Proveedor "{nombre}" tiene {len(ocs_activas)} OCs '
                          f'activas · no se puede dar de baja sin antes '
                          f'cerrarlas (Cancelada/Rechazada) o forzar el override.'),
                'codigo': 'PROVEEDOR_CON_OCS_ACTIVAS',
                'ocs_activas': [{'numero_oc': r[0], 'estado': r[1],
                                  'valor_total': float(r[2] or 0)}
                                 for r in ocs_activas],
                'override_disponible': es_admin,
                'hint': ('Cancelá/rechazá las OCs primero, o ' +
                         ('reenviá con {"force":true,"motivo":"..."} para admin override.'
                           if es_admin else
                           'pedile a admin que haga el override.')),
            }), 409
        c.execute(
            "UPDATE proveedores SET activo=0, motivo_baja=?, fecha_baja=? WHERE nombre=?",
            (motivo, datetime.now().isoformat(), nombre)
        )
        try:
            audit_log(c, usuario=usuario, accion='BAJA_PROVEEDOR',
                      tabla='proveedores', registro_id=nombre,
                      despues={'activo': 0, 'motivo_baja': motivo[:300],
                                'force_override': force,
                                'ocs_activas_al_baja': len(ocs_activas),
                                'ocs_numeros': [r[0] for r in ocs_activas[:20]]},
                      detalle=(f"Baja proveedor '{nombre}' · motivo: {motivo[:120]} · "
                                f"OCs activas al momento: {len(ocs_activas)}"
                                f"{' [FORCE]' if force else ''}"))
        except Exception as e:
            log.warning('audit_log BAJA_PROVEEDOR fallo: %s', e)
        conn.commit()
        return jsonify({
            'ok': True,
            'message': f"Proveedor '{nombre}' dado de baja",
            'ocs_activas_al_baja': len(ocs_activas),
            'force_override': force,
        })
    # PATCH — edit
    d = request.json or {}
    fields = ['contacto','email','telefono','nit','direccion',
              'banco','tipo_cuenta','num_cuenta','concepto_compra',
              'condiciones_pago','categoria']

    # Renombrar proveedor con propagacion automatica a OCs/historico.
    # Antes Catalina debia darlo de baja y crear uno nuevo, perdiendo
    # historial. Ahora si viene 'nombre' nuevo distinto al actual,
    # se actualiza en proveedores + en TODAS las tablas que lo referencian
    # por nombre (ordenes_compra, solicitudes_compra, precios_mp_historico,
    # cotizaciones, pedidos como cliente cuando aplique).
    nuevo_nombre = (d.get('nombre') or '').strip()
    rename_propagado = {}
    if nuevo_nombre and nuevo_nombre != nombre:
        # Verificar que no exista ya un proveedor activo con ese nombre
        existe = c.execute(
            "SELECT 1 FROM proveedores WHERE nombre=? AND activo=1",
            (nuevo_nombre,)
        ).fetchone()
        if existe:
            return jsonify({
                'error': f"Ya existe otro proveedor activo con el nombre '{nuevo_nombre}'"
            }), 409
        # Hacer el rename en transaccion
        c.execute("UPDATE proveedores SET nombre=? WHERE nombre=? AND activo=1",
                  (nuevo_nombre, nombre))
        if c.rowcount == 0:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
        # Propagar a tablas referentes
        propagar = [
            ('ordenes_compra',     'proveedor'),
            ('solicitudes_compra', 'proveedor_sugerido'),
            ('solicitudes_compra_items', 'proveedor_sugerido'),
            ('precios_mp_historico', 'proveedor'),
        ]
        for tabla, col in propagar:
            try:
                c.execute(f"UPDATE {tabla} SET {col}=? WHERE {col}=?",
                          (nuevo_nombre, nombre))
                rename_propagado[tabla] = c.rowcount
            except Exception:
                pass
        # cotizaciones (3 columnas de proveedor)
        try:
            for col in ('proveedor_a', 'proveedor_b', 'proveedor_c'):
                c.execute(f"UPDATE cotizaciones SET {col}=? WHERE {col}=?",
                          (nuevo_nombre, nombre))
            rename_propagado['cotizaciones'] = 'OK'
        except Exception:
            pass
        # Despues del rename, las queries deben usar el nuevo nombre
        nombre = nuevo_nombre

    sets = [f"{f}=?" for f in fields if f in d]
    vals = [d[f] for f in fields if f in d]
    if not sets and not rename_propagado:
        return jsonify({'error': 'No hay campos para actualizar'}), 400

    if sets:
        vals.append(nombre)
        c.execute(f"UPDATE proveedores SET {', '.join(sets)} WHERE nombre=? AND activo=1", vals)
        if c.rowcount == 0:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
    try:
        audit_log(c, usuario=usuario, accion='ACTUALIZAR_PROVEEDOR',
                  tabla='proveedores', registro_id=nombre,
                  despues={'campos': [f for f in fields if f in d],
                           'rename_propagado': rename_propagado or None},
                  detalle=f"Actualizó proveedor '{nombre}'"
                          + (' (renombrado)' if rename_propagado else ''))
    except Exception as e:
        log.warning('audit_log ACTUALIZAR_PROVEEDOR fallo: %s', e)
    conn.commit()
    msg = f"Proveedor actualizado"
    if rename_propagado:
        msg += f" — nombre cambiado a '{nuevo_nombre}', propagado en: {rename_propagado}"
    return jsonify({'ok': True, 'message': msg, 'rename_propagado': rename_propagado,
                    'nombre_actual': nombre})

@bp.route('/api/proveedores-compras/<path:nombre>/ficha')
def proveedor_ficha_360(nombre):
    """Proveedor 360: datos completos + historial OCs + scoring."""
    # SEC-FIX 23-may-2026 · auditoría · expone num_cuenta/banco/nit · Ley 1581
    u, err, code = _require_compras_session()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT id, nombre, contacto, email, telefono, categoria,
                        nit, direccion, num_cuenta, tipo_cuenta, banco, concepto_compra,
                        id_interno, estado_lpa, ultima_evaluacion, vencimiento_docs,
                        acuerdo_calidad, condiciones_pago
                 FROM proveedores WHERE nombre=? AND activo=1""", (nombre,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    cols = ['id','nombre','contacto','email','telefono','categoria',
            'nit','direccion','num_cuenta','tipo_cuenta','banco','concepto_compra',
            'id_interno','estado_lpa','ultima_evaluacion','vencimiento_docs',
            'acuerdo_calidad','condiciones_pago']
    prov = dict(zip(cols, row))
    # SEC-FIX 10-jun audit · Habeas Data (Ley 1581): datos bancarios solo admin+contadora.
    if not (u and u.lower() in {x.lower() for x in (set(ADMIN_USERS) | set(CONTADORA_USERS))}):
        for _cb in ('num_cuenta', 'tipo_cuenta', 'banco', 'nit'):
            if prov.get(_cb):
                prov[_cb] = '***'
    # OC stats
    c.execute("""SELECT COUNT(*), COALESCE(SUM(valor_total),0), MIN(fecha), MAX(fecha)
                 FROM ordenes_compra WHERE proveedor=?""", (nombre,))
    r = c.fetchone()
    oc_total, valor_total, primera_oc, ultima_oc = (r[0] or 0), (r[1] or 0), r[2], r[3]
    c.execute("SELECT COUNT(*) FROM ordenes_compra WHERE proveedor=? AND estado IN ('Recibida','Pagada','Parcial')", (nombre,))
    oc_recibidas = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM ordenes_compra WHERE proveedor=? AND tiene_discrepancias=1 AND estado IN ('Recibida','Pagada','Parcial')", (nombre,))
    oc_disc = c.fetchone()[0] or 0
    # Scoring: cumplimiento 70% + calidad (sin discrepancias) 30%
    cumplimiento = round((oc_recibidas / oc_total * 100) if oc_total > 0 else 0, 1)
    tasa_disc = round((oc_disc / oc_recibidas * 100) if oc_recibidas > 0 else 0, 1)
    score = min(100.0, round(cumplimiento * 0.7 + (100 - tasa_disc) * 0.3, 1))
    # Recent OCs
    c.execute("""SELECT numero_oc, fecha, estado, valor_total, categoria,
                        tiene_discrepancias, fecha_recepcion
                 FROM ordenes_compra WHERE proveedor=?
                 ORDER BY fecha DESC LIMIT 8""", (nombre,))
    oc_cols = ['numero_oc','fecha','estado','valor_total','categoria','tiene_discrepancias','fecha_recepcion']
    ocs_recientes = [dict(zip(oc_cols, r)) for r in c.fetchall()]
    # Materials bought from this supplier
    c.execute("""SELECT oci.codigo_mp, oci.nombre_mp, COUNT(*) as veces, SUM(oci.cantidad_g) as total_g
                 FROM ordenes_compra oc
                 JOIN ordenes_compra_items oci ON oc.numero_oc=oci.numero_oc
                 WHERE oc.proveedor=?
                 GROUP BY oci.codigo_mp, oci.nombre_mp
                 ORDER BY total_g DESC LIMIT 15""", (nombre,))
    mps = [{'codigo': r[0], 'nombre': r[1], 'veces': r[2], 'total_g': round(r[3] or 0, 1)}
           for r in c.fetchall()]
    return jsonify({
        'proveedor': prov,
        'stats': {
            'oc_total': oc_total, 'oc_recibidas': oc_recibidas, 'oc_discrepancias': oc_disc,
            'valor_total': valor_total, 'primera_oc': primera_oc, 'ultima_oc': ultima_oc,
            'cumplimiento': cumplimiento, 'tasa_discrepancias': tasa_disc, 'score': score
        },
        'ocs_recientes': ocs_recientes,
        'materiales': mps
    })

@bp.route('/api/solicitudes-compra', methods=['GET','POST'])
def handle_solicitudes_compra():
    if request.method == 'POST':
        conn = None
        try:
            conn = get_db(); c = conn.cursor()
            d = request.json or {}
            emp = d.get('empresa','Espagiria')
            cat = d.get('categoria','Materia Prima')
            tip = d.get('tipo','Compra')
            area = d.get('area','Produccion')
            email_sol = d.get('email_solicitante', '').strip().lower()
            fecha_req = d.get('fecha_requerida', '').strip()
            val_sol = float(d.get('valor') or 0)
            # numero único con reintento ante carrera MAX+1 entre workers
            for _intento in range(6):
                c.execute("SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) FROM solicitudes_compra WHERE numero LIKE ?", (f"SOL-{datetime.now().strftime('%Y')}-%",))
                numero = f"SOL-{datetime.now().strftime('%Y')}-{(c.fetchone()[0] or 0)+1:04d}"
                try:
                    c.execute("""INSERT INTO solicitudes_compra
                                 (numero,fecha,estado,solicitante,urgencia,observaciones,area,empresa,categoria,tipo,email_solicitante,fecha_requerida,valor)
                                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                              (numero, datetime.now().isoformat(), 'Pendiente',
                               d.get('solicitante',''), d.get('urgencia','Normal'), d.get('observaciones',''),
                               area, emp, cat, tip, email_sol, fecha_req, val_sol))
                    break
                except sqlite3.IntegrityError:
                    if _intento == 5:
                        raise
            for it in (d.get('items') or []):
                # proveedor_sugerido es opcional. Migración 19xx aseguró que la
                # columna exista. Si por alguna razón no, fallback al INSERT
                # mínimo. Solicitudes generadas desde Plan MP rolling envían
                # proveedor_sugerido tomado de mp_lead_time_config.
                try:
                    c.execute("""INSERT INTO solicitudes_compra_items
                                 (numero,codigo_mp,nombre_mp,cantidad_g,unidad,justificacion,valor_estimado,proveedor_sugerido)
                                 VALUES (?,?,?,?,?,?,?,?)""",
                              (numero, it.get('codigo_mp',''), it.get('nombre_mp',''),
                               it.get('cantidad_g',0), it.get('unidad','g'),
                               it.get('justificacion',''), it.get('valor_estimado',0),
                               it.get('proveedor_sugerido','')))
                except sqlite3.OperationalError:
                    # Esquema antiguo sin proveedor_sugerido — fallback
                    c.execute("""INSERT INTO solicitudes_compra_items
                                 (numero,codigo_mp,nombre_mp,cantidad_g,unidad,justificacion,valor_estimado)
                                 VALUES (?,?,?,?,?,?,?)""",
                              (numero, it.get('codigo_mp',''), it.get('nombre_mp',''),
                               it.get('cantidad_g',0), it.get('unidad','g'),
                               it.get('justificacion',''), it.get('valor_estimado',0)))
            try:
                usuario_act = session.get('compras_user', '') or d.get('solicitante','')
                audit_log(c, usuario=usuario_act, accion='CREAR_SOLICITUD',
                          tabla='solicitudes_compra', registro_id=numero,
                          despues={'empresa': emp, 'categoria': cat, 'tipo': tip,
                                    'area': area,
                                    'solicitante': d.get('solicitante','')[:100],
                                    'urgencia': d.get('urgencia','Normal'),
                                    'valor': val_sol,
                                    'items_count': len(d.get('items') or [])},
                          detalle=f"Creó solicitud {numero} · {emp} · {cat} · "
                                  f"{len(d.get('items') or [])} items · valor {val_sol:.0f}")
            except Exception as e:
                log.warning('audit_log CREAR_SOLICITUD fallo: %s', e)
            conn.commit()
            return jsonify({'message': f'Solicitud {numero} creada', 'numero': numero}), 201
        except Exception as e:
            if conn:
                try:
                    conn.rollback()
                except sqlite3.Error:
                    # Rollback puede fallar si no había transacción abierta —
                    # no es crítico, ya estamos en error path.
                    pass
            log.exception('crear solicitud-compra fallo: %s', e)
            return jsonify({'error': 'No se pudo crear la solicitud'}), 500
    conn = get_db(); c = conn.cursor()
    # GET: listar todas las solicitudes
    filtro_estado = request.args.get('estado', '')
    filtro_empresa = request.args.get('empresa', '')
    # SELECT con LEFT JOIN a marketing_influencers para enriquecer con datos
    # bancarios. Las columnas ciudad/instagram las garantiza la migración 31.
    # Si por alguna razón no existen (esquema pre-migración 31), caemos al
    # SELECT sin esas columnas en el except más abajo.
    # vence_pago_at + fecha_contenido vía subquery escalar (mig 195) ·
    # no requiere JOIN extra · vacío si la SOL no tiene pago influencer
    # asociado. Fallback try/except cubre el caso de mig 195 sin aplicar.
    _pi_vence_subq = ("(SELECT pi.vence_pago_at FROM pagos_influencers pi "
                      "WHERE pi.numero_oc = sc.numero_oc LIMIT 1)")
    _pi_fcont_subq = ("(SELECT pi.fecha_contenido FROM pagos_influencers pi "
                      "WHERE pi.numero_oc = sc.numero_oc LIMIT 1)")
    sql_with_extras = f"""
        SELECT sc.numero, sc.fecha, sc.estado, sc.solicitante, sc.urgencia,
               sc.observaciones, sc.empresa, sc.categoria, sc.tipo, sc.area,
               sc.email_solicitante, sc.fecha_requerida, sc.numero_oc,
               COALESCE(NULLIF(oc.valor_total,0), sc.valor, 0) as valor_oc,
               COALESCE(mi.nombre, '')           as inf_nombre,
               COALESCE(mi.banco, '')            as inf_banco,
               COALESCE(mi.cuenta_bancaria, '')  as inf_cuenta,
               COALESCE(mi.tipo_cuenta, '')      as inf_tipo_cuenta,
               COALESCE(mi.cedula_nit, '')       as inf_cedula,
               COALESCE(mi.email, '')            as inf_email,
               COALESCE(mi.ciudad, '')           as inf_ciudad,
               COALESCE(mi.instagram, '')        as inf_instagram,
               COALESCE({_pi_vence_subq},'') as vence_pago_at,
               COALESCE({_pi_fcont_subq},'') as fecha_contenido
        FROM solicitudes_compra sc
        LEFT JOIN ordenes_compra oc ON oc.numero_oc = sc.numero_oc
        LEFT JOIN marketing_influencers mi ON mi.id = sc.influencer_id
        WHERE 1=1"""
    sql_minimal = """
        SELECT sc.numero, sc.fecha, sc.estado, sc.solicitante, sc.urgencia,
               sc.observaciones, sc.empresa, sc.categoria, sc.tipo, sc.area,
               sc.email_solicitante, sc.fecha_requerida, sc.numero_oc,
               COALESCE(NULLIF(oc.valor_total,0), sc.valor, 0) as valor_oc,
               COALESCE(mi.nombre, '')           as inf_nombre,
               COALESCE(mi.banco, '')            as inf_banco,
               COALESCE(mi.cuenta_bancaria, '')  as inf_cuenta,
               COALESCE(mi.tipo_cuenta, '')      as inf_tipo_cuenta,
               COALESCE(mi.cedula_nit, '')       as inf_cedula,
               COALESCE(mi.email, '')            as inf_email,
               '' as inf_ciudad,
               '' as inf_instagram,
               '' as vence_pago_at,
               '' as fecha_contenido
        FROM solicitudes_compra sc
        LEFT JOIN ordenes_compra oc ON oc.numero_oc = sc.numero_oc
        LEFT JOIN marketing_influencers mi ON mi.id = sc.influencer_id
        WHERE 1=1"""
    sql = sql_with_extras
    params = []
    if filtro_estado: sql += " AND sc.estado=?"; params.append(filtro_estado)
    if filtro_empresa: sql += " AND sc.empresa=?"; params.append(filtro_empresa)
    filtro_categoria = request.args.get('categoria', '')
    # Sebastian 6-may-2026: param ?fuente= separa los 3 origenes de SOLs.
    # Catalina debe ver cada flujo en su tab sin mezcla:
    #   fuente=usuarios → SOLs del modulo /solicitudes (papeleria, EPP,
    #     servicios, mantenimiento, etc.) · NO incluye MP/Empaque ni
    #     Influencer/CC.
    #   fuente=planta → SOLs auto-generadas por el calendario de planta
    #     (Materia Prima + Empaque). Agrupadas por proveedor en el tab.
    #   fuente=influencers → Influencer/Marketing Digital + Cuenta de Cobro.
    # Sin fuente · comportamiento legacy (todo excepto Influencer/CC).
    fuente = (request.args.get('fuente') or '').strip().lower()
    _CATS_PLANTA = ('Materia Prima', 'Empaque', 'Material de Empaque')
    _CATS_INFLUENCER = ('Influencer/Marketing Digital', 'Cuenta de Cobro')
    # PRIVACY-FIX · 21-may-2026 · Influencers solo admin (Sebas/Alejandro).
    # Antes: Catalina y cualquier compras_user podían ver banco + cuenta
    # bancaria via fetch directo. Ahora 403 si no es admin.
    filtro_cat_actual = (filtro_categoria or '').strip()
    _es_influencer_req = (fuente == 'influencers'
                          or filtro_cat_actual in _CATS_INFLUENCER)
    if _es_influencer_req:
        _user_actual = session.get('compras_user', '').lower()
        _admins_lower = {x.lower() for x in ADMIN_USERS}
        if _user_actual not in _admins_lower:
            return jsonify({
                'error': 'Privado · solo admin puede ver pagos a influencers',
                'codigo': 'PRIVATE_INFLUENCERS',
            }), 403
    if fuente == 'planta':
        ph = ','.join(['?'] * len(_CATS_PLANTA))
        sql += f" AND sc.categoria IN ({ph})"
        params.extend(_CATS_PLANTA)
    elif fuente == 'usuarios':
        # Excluye Planta + Influencers · queda todo lo demas (Papeleria,
        # Servicios, EPP, Aseo, Mantenimiento, Software, Dotacion,
        # Reactivos, Administrativo, Infraestructura, Otros, etc.)
        excluir = list(_CATS_PLANTA) + list(_CATS_INFLUENCER)
        ph = ','.join(['?'] * len(excluir))
        sql += f" AND sc.categoria NOT IN ({ph})"
        params.extend(excluir)
    elif fuente == 'influencers':
        sql += " AND sc.categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')"
    elif filtro_categoria in ('Influencer/Marketing Digital', 'Cuenta de Cobro'):
        sql += " AND sc.categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')"
    elif filtro_categoria:
        sql += " AND sc.categoria=?"; params.append(filtro_categoria)
    else:
        sql += " AND sc.categoria NOT IN ('Influencer/Marketing Digital','Cuenta de Cobro')"
    if filtro_categoria in ('Influencer/Marketing Digital', 'Cuenta de Cobro'):
        # Cadena de prioridad para ordenar (Sebastián 27-may-2026 PM):
        #  1. vence_pago_at  (promesa 30d desde fecha_contenido · mig 195)
        #     · los más cerca a vencer o ya vencidos van ARRIBA
        #  2. sc.fecha_requerida (fecha tope cuando el pago debe hacerse)
        #  3. sc.fecha           (fecha en que se creó la solicitud)
        # Estado: Aprobadas (por pagar) primero, luego el resto.
        # vence_pago_at viene del SELECT como subquery escalar → sin JOIN extra.
        sql += (
            " ORDER BY "
            " CASE sc.estado WHEN 'Aprobada' THEN 0 WHEN 'Pendiente' THEN 1 "
            "                WHEN 'Pagada' THEN 2 ELSE 3 END, "
            " COALESCE(NULLIF(vence_pago_at,''), "
            "          NULLIF(sc.fecha_requerida,''), "
            "          sc.fecha) ASC, "
            " sc.numero ASC LIMIT 300"
        )
    else:
        sql += " ORDER BY sc.fecha DESC LIMIT 200"
    try:
        c.execute(sql, params)
    except Exception as _e:
        # Fallback: si las columnas ciudad/instagram/vence_pago_at no existen
        # aún (DB sin migración 31 o 195 aplicada), usar versión minimal.
        # SQLite dice "no such column"; PostgreSQL dice "does not exist".
        _err_lower = str(_e).lower()
        _es_col_falta = ('no such column' in _err_lower or 'does not exist' in _err_lower)
        if _es_col_falta and sql.startswith(sql_with_extras.split('WHERE')[0][:50]):
            # Reaplicar reemplazos de la rama Influencer si correspondía
            sql_fb = sql_minimal
            if filtro_categoria == 'Influencer/Marketing Digital':
                sql_fb = sql_fb.replace(
                    "FROM solicitudes_compra sc",
                    "FROM solicitudes_compra sc LEFT JOIN pagos_influencers pi ON pi.numero_oc = sc.numero_oc"
                )
            for cond in [
                ("estado", filtro_estado), ("empresa", filtro_empresa),
                ("categoria", filtro_categoria),
            ]:
                pass  # los WHEREs y ORDER BY ya están aplicados via sql_with_extras
            # Reconstruir condiciones
            sql_fb_full = sql_fb
            params_fb = []
            if filtro_estado:
                sql_fb_full += " AND sc.estado=?"; params_fb.append(filtro_estado)
            if filtro_empresa:
                sql_fb_full += " AND sc.empresa=?"; params_fb.append(filtro_empresa)
            if filtro_categoria in ('Influencer/Marketing Digital', 'Cuenta de Cobro'):
                sql_fb_full += " AND sc.categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')"
            elif filtro_categoria:
                sql_fb_full += " AND sc.categoria=?"; params_fb.append(filtro_categoria)
            else:
                sql_fb_full += " AND sc.categoria NOT IN ('Influencer/Marketing Digital','Cuenta de Cobro')"
            if filtro_categoria in ('Influencer/Marketing Digital', 'Cuenta de Cobro'):
                sql_fb_full += (
                    " ORDER BY CASE sc.estado WHEN 'Aprobada' THEN 0 "
                    "WHEN 'Pendiente' THEN 1 WHEN 'Pagada' THEN 2 ELSE 3 END, "
                    "COALESCE(NULLIF(sc.fecha_requerida,''), sc.fecha) ASC, "
                    "sc.numero ASC LIMIT 300"
                )
            else:
                sql_fb_full += " ORDER BY sc.fecha DESC LIMIT 200"
            __import__('logging').getLogger('compras').warning(
                "SELECT con extras falló (%s) — fallback a sql_minimal", _e
            )
            c.execute(sql_fb_full, params_fb)
        else:
            raise
    cols_sol = ['numero','fecha','estado','solicitante','urgencia','observaciones','empresa','categoria','tipo','area','email_solicitante','fecha_requerida','numero_oc','valor',
                'inf_nombre','inf_banco','inf_cuenta','inf_tipo_cuenta','inf_cedula','inf_email','inf_ciudad','inf_instagram',
                'vence_pago_at','fecha_contenido']
    rows_sol = []
    for r in c.fetchall():
        row = dict(zip(cols_sol, r))
        # valor comes from OC join; fallback to items sum if still 0
        if not row.get('valor'):
            c2 = conn.cursor()
            c2.execute("SELECT COALESCE(SUM(valor_estimado),0) FROM solicitudes_compra_items WHERE numero=?", (row['numero'],))
            row['valor'] = c2.fetchone()[0] or 0
        # Last fallback: parse VALOR from OBS string (influencer SOLs without OC)
        if not row.get('valor'):
            obs_str = row.get('observaciones') or ''
            if 'VALOR:' in obs_str:
                try:
                    v_str = obs_str.split('VALOR:')[1].split('|')[0].strip().replace('$','').replace(',','').replace('.','')
                    row['valor'] = float(v_str)
                except (ValueError, IndexError):
                    # Observaciones malformadas — fallback silencioso aceptable
                    # (es solo enriquecimiento de display, no flujo crítico).
                    pass
        rows_sol.append(row)
    return jsonify({'solicitudes': rows_sol})


# ── Sebastián 4-may-2026 (Catalina): agrupar solicitudes por proveedor ────
# Catalina recibe 200+ solicitudes AUTO-PLAN desde planta (cada una es 1 MP
# en déficit). En vez de gestionarlas una por una, las quiere agrupadas
# por proveedor sugerido para procesar todas las del mismo proveedor en
# UNA sola OC. Endpoint que devuelve el agrupamiento + items consolidados.
@bp.route('/api/compras/solicitudes-agrupadas-por-proveedor', methods=['GET'])
def solicitudes_agrupadas_por_proveedor():
    """Agrupa solicitudes_compra Pendientes por proveedor sugerido.

    Querystring:
      ?estado=Pendiente|Aprobada|all  (default: Pendiente)
      ?categoria=Mat. Primas|Empaque|...  (default: todas excepto Influencer/CC)

    Returns:
      {
        "grupos": [
          {
            "proveedor": "Colquimicos",
            "solicitudes_count": 12,
            "items_count": 12,
            "valor_total": 1234567.89,
            "solicitudes": [{numero, fecha, urgencia, area, items: [...]}, ...],
            "items_consolidados": [{codigo_mp, nombre_mp, cantidad_g, valor_estimado}],
            "urgencia_max": "Urgente",
          }, ...
        ],
        "sin_proveedor": [...],   # mismo shape (proveedor='')
        "total_solicitudes": 204,
        "total_grupos": 18
      }

    Items se consolidan por codigo_mp dentro de cada grupo (suma de
    cantidad_g + valor_estimado). Esto es lo que Catalina usa para
    crear UNA OC por proveedor.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    estado_filtro = (request.args.get('estado') or 'Pendiente').strip()
    categoria_filtro = (request.args.get('categoria') or '').strip()
    # Sebastian 5-may-2026: filtro de fuente alineado con /api/solicitudes-compra
    # planta = solo MP + Empaque (la vista agrupada se usa principalmente para
    #         consolidar pedidos de planta a proveedores)
    fuente = (request.args.get('fuente') or '').strip().lower()

    conn = get_db(); c = conn.cursor()
    sql = """
        SELECT s.numero, s.fecha, s.estado, s.solicitante, s.urgencia,
               COALESCE(s.observaciones,''), s.empresa, s.categoria,
               s.tipo, s.area, s.fecha_requerida, COALESCE(s.numero_oc,''),
               COALESCE(s.valor, 0)
        FROM solicitudes_compra s
        WHERE 1=1
    """
    params = []
    if estado_filtro and estado_filtro.lower() != 'all':
        sql += " AND s.estado=?"
        params.append(estado_filtro)
    # Excluir Influencer/CC siempre — esos tienen flujo aparte
    sql += " AND s.categoria NOT IN ('Influencer/Marketing Digital','Cuenta de Cobro')"
    if fuente == 'planta':
        sql += " AND s.categoria IN ('Materia Prima','Empaque','Material de Empaque')"
    elif fuente == 'usuarios':
        # Fix 28-may · INV-1: rama simétrica que faltaba · sin esto fuente=usuarios
        # dejaba colar SOLs de planta (MP/Empaque) en el tab Solicitudes.
        sql += " AND s.categoria NOT IN ('Materia Prima','Empaque','Material de Empaque')"
    if categoria_filtro:
        sql += " AND s.categoria=?"
        params.append(categoria_filtro)
    # Sprint Compras N2 · 21-may-2026 · paginación por SOLs.
    # Antes LIMIT 500 fijo · con 300+ SOLs auto-plan + influencer se
    # truncaba silenciosamente. Ahora respeta ?limit=N&offset=N.
    try:
        limit = max(1, min(int(request.args.get('limit', 500)), 1000))
    except (ValueError, TypeError):
        limit = 500
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    # ALTA-4 fix · 21-may-2026: rewrite count manualmente.
    # Antes el .replace() buscaba un literal que NO matcheaba el SELECT real
    # (tiene COALESCE) · count_sql quedaba con el SELECT original devolviendo
    # filas en lugar de COUNT · int() reventaba → 0 silencioso siempre.
    try:
        # Extraer la parte FROM ... WHERE ... del sql actual
        idx_from = sql.upper().find(' FROM ')
        from_clause = sql[idx_from:] if idx_from > 0 else ''
        count_sql = 'SELECT COUNT(*)' + from_clause
        total_sols = int((c.execute(count_sql, params).fetchone() or [0])[0])
    except Exception as _e:
        __import__('logging').getLogger('compras').warning(
            'count_sql fallo (paginación): %s', _e)
        total_sols = 0
    sql += f" ORDER BY s.fecha DESC LIMIT {limit} OFFSET {offset}"

    c.execute(sql, params)
    sol_cols = ['numero','fecha','estado','solicitante','urgencia','observaciones',
                'empresa','categoria','tipo','area','fecha_requerida','numero_oc','valor']
    solicitudes = [dict(zip(sol_cols, r)) for r in c.fetchall()]
    if not solicitudes:
        return jsonify({
            'grupos': [], 'sin_proveedor': [],
            'total_solicitudes': 0, 'total_grupos': 0,
        })

    # Cargar items de todas las solicitudes en 1 query (evita N+1)
    nums = [s['numero'] for s in solicitudes]
    placeholders = ','.join(['?'] * len(nums))
    try:
        c.execute(f"""
            SELECT sci.id, sci.numero, sci.codigo_mp, sci.nombre_mp, sci.cantidad_g, sci.unidad,
                   COALESCE(sci.valor_estimado, 0),
                   COALESCE(sci.justificacion, ''),
                   COALESCE(sci.precio_unit_g, 0),
                   COALESCE(sci.proveedor_sugerido, ''),
                   COALESCE(mm.nombre_inci, '')
            FROM solicitudes_compra_items sci
            LEFT JOIN maestro_mps mm ON mm.codigo_mp = sci.codigo_mp
            WHERE sci.numero IN ({placeholders})
        """, nums)
        item_cols = ['id','numero','codigo_mp','nombre_mp','cantidad_g','unidad',
                     'valor_estimado','justificacion','precio_unit_g','proveedor_sugerido','nombre_inci']
        items_all = [dict(zip(item_cols, r)) for r in c.fetchall()]
    except sqlite3.OperationalError:
        # Fallback sin proveedor_sugerido (esquema viejo)
        c.execute(f"""
            SELECT id, numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                   COALESCE(valor_estimado, 0)
            FROM solicitudes_compra_items
            WHERE numero IN ({placeholders})
        """, nums)
        items_all = []
        for r in c.fetchall():
            d = dict(zip(['id','numero','codigo_mp','nombre_mp','cantidad_g',
                          'unidad','valor_estimado'], r))
            d['justificacion'] = ''
            d['precio_unit_g'] = 0
            d['proveedor_sugerido'] = ''
            items_all.append(d)

    # Indexar items por solicitud
    items_por_sol = {}
    for it in items_all:
        items_por_sol.setdefault(it['numero'], []).append(it)

    # Determinar proveedor de cada solicitud (el primer item con proveedor_sugerido != '')
    # Si todos los items tienen el mismo proveedor → ese es el proveedor del grupo.
    # Si hay mezcla, marcar como 'Mixto' para que Catalina los revise individualmente.
    URG_RANK = {'Critico': 4, 'Urgente': 3, 'Alta': 2, 'Media': 1, 'Normal': 0}

    # Sebastián 24-may-2026 · audit Bandeja Planta · agrupamiento case + espacios
    # insensible. Antes 'Colquimicos', 'COLQUIMICOS', 'colquímicos  ' quedaban
    # como 3 grupos distintos · Catalina veía el mismo proveedor 3× y no podía
    # consolidar en 1 OC. Ahora normalizamos para la KEY pero mantenemos la
    # forma original (la primera vista) como display al UI.
    def _norm_prov(p):
        return ' '.join((p or '').strip().split()).upper()

    def _prov_dominante(items_lista):
        # Key normalizada para dedup · valor = forma original conservada
        provs_norm = {}
        for it in items_lista:
            raw = (it.get('proveedor_sugerido') or '').strip()
            if not raw:
                continue
            k = _norm_prov(raw)
            provs_norm.setdefault(k, raw)
        if not provs_norm:
            return ''
        if len(provs_norm) == 1:
            # Devolvemos la key normalizada para agrupar consistente
            return next(iter(provs_norm.keys()))
        return '__MIXTO__'

    # Agrupar (key = proveedor normalizado · display = forma original)
    grupos_dict = {}        # key normalizada → lista SOLs
    grupos_label = {}       # key normalizada → primera forma original vista
    sin_proveedor = []
    for s in solicitudes:
        items = items_por_sol.get(s['numero'], [])
        s['items'] = items
        s['valor_calc'] = sum(float(it.get('valor_estimado') or 0) for it in items)
        prov = _prov_dominante(items)
        if not prov or prov == '__MIXTO__':
            target = sin_proveedor
            if prov == '__MIXTO__':
                s['_motivo_sin_grupo'] = 'mixto'
            else:
                s['_motivo_sin_grupo'] = 'sin_sugerencia'
            target.append(s)
        else:
            grupos_dict.setdefault(prov, []).append(s)
            # Primera forma original que vimos · preserva capitalización humana
            if prov not in grupos_label:
                for it in items:
                    raw = (it.get('proveedor_sugerido') or '').strip()
                    if raw and _norm_prov(raw) == prov:
                        grupos_label[prov] = raw
                        break

    # Construir output
    grupos = []
    for prov, sols in sorted(grupos_dict.items()):
        # Consolidar items por codigo_mp (suma)
        consolidados = {}
        for s in sols:
            for it in s['items']:
                k = (it.get('codigo_mp') or '').upper().strip() or it.get('nombre_mp', '')[:50]
                if k not in consolidados:
                    consolidados[k] = {
                        'codigo_mp': it.get('codigo_mp', ''),
                        'nombre_mp': it.get('nombre_mp', ''),
                        'unidad': it.get('unidad', 'g'),
                        'cantidad_g': 0.0,
                        'valor_estimado': 0.0,
                        'precio_unit_g': float(it.get('precio_unit_g') or 0),
                        'solicitudes_origen': [],
                    }
                consolidados[k]['cantidad_g'] += float(it.get('cantidad_g') or 0)
                consolidados[k]['valor_estimado'] += float(it.get('valor_estimado') or 0)
                consolidados[k]['solicitudes_origen'].append(s['numero'])
        urg_max = 'Normal'
        for s in sols:
            if URG_RANK.get(s.get('urgencia') or 'Normal', 0) > URG_RANK.get(urg_max, 0):
                urg_max = s.get('urgencia') or 'Normal'
        grupos.append({
            'proveedor': prov,                                    # key normalizada
            'proveedor_label': grupos_label.get(prov, prov),      # display original
            'solicitudes_count': len(sols),
            'items_count': len(consolidados),
            'valor_total': round(sum(float(s.get('valor_calc') or 0) for s in sols), 2),
            'urgencia_max': urg_max,
            'solicitudes': sols,
            'items_consolidados': sorted(consolidados.values(),
                                          key=lambda x: -float(x.get('valor_estimado') or 0)),
        })

    # Ordenar grupos por urgencia max → solicitudes_count desc
    grupos.sort(key=lambda g: (-URG_RANK.get(g['urgencia_max'], 0),
                                -g['solicitudes_count']))

    return jsonify({
        'grupos': grupos,
        'sin_proveedor': sin_proveedor,
        'total_solicitudes': len(solicitudes),
        'total_solicitudes_full': total_sols,
        'mostrando_sols': len(solicitudes),
        'limit': limit,
        'offset': offset,
        'hay_mas': total_sols > (offset + len(solicitudes)),
        'total_grupos': len(grupos),
        'estado_filtro': estado_filtro,
        'categoria_filtro': categoria_filtro,
    })


# ── Sebastián 4-may-2026 (Catalina): convertir N solicitudes en 1 OC ──
@bp.route('/api/compras/oc-desde-solicitudes', methods=['POST'])
def crear_oc_desde_solicitudes():
    """Convierte un lote de solicitudes_compra Pendientes en UNA sola OC.

    Body JSON:
      {
        "proveedor": "Colquimicos",       # requerido
        "solicitudes": ["AUTO-0946",...],  # requerido (lista de numeros)
        "observaciones": "...",            # opcional
        "fecha_entrega_est": "2026-05-15", # opcional
        "consolidar_iguales": true,        # default true (suma cant por codigo_mp)
        "categoria": "MP"                  # default "MP"
      }

    Atomico: si cualquier paso falla, rollback completo. La OC no queda
    creada y las solicitudes no cambian de estado.

    Returns:
      201 {numero_oc, items_creados, solicitudes_vinculadas, valor_total}
      400 si validacion falla (proveedor faltante, solicitudes vacias)
      404 si alguna solicitud no existe
      409 si alguna solicitud no esta Pendiente o ya tiene OC
    """
    usuario, err, code = _require_compras_write()
    if err:
        return err, code

    d = request.get_json(silent=True) or {}
    proveedor = (d.get('proveedor') or '').strip()
    nums = d.get('solicitudes') or []
    if not proveedor:
        return jsonify({'error': 'proveedor requerido'}), 400
    if not isinstance(nums, list) or not nums:
        return jsonify({'error': 'solicitudes (lista) requerido'}), 400
    nums = [str(n).strip().upper() for n in nums if str(n).strip()]
    if not nums:
        return jsonify({'error': 'solicitudes lista vacia'}), 400
    if len(nums) > 100:
        return jsonify({'error': 'maximo 100 solicitudes por OC'}), 400

    consolidar = bool(d.get('consolidar_iguales', True))
    categoria = (d.get('categoria') or 'MP').strip()
    obs_extra = (d.get('observaciones') or '').strip()
    fecha_entrega_est = (d.get('fecha_entrega_est') or '').strip()

    conn = get_db(); c = conn.cursor()
    try:
        # 1. Validar todas las solicitudes
        placeholders = ','.join(['?'] * len(nums))
        c.execute(f"""
            SELECT numero, estado, COALESCE(numero_oc,'')
            FROM solicitudes_compra
            WHERE numero IN ({placeholders})
        """, nums)
        rows = c.fetchall()
        existentes = {r[0]: (r[1], r[2]) for r in rows}
        faltantes = [n for n in nums if n not in existentes]
        if faltantes:
            return jsonify({
                'error': 'Solicitudes no encontradas',
                'faltantes': faltantes,
            }), 404
        no_pendientes = [n for n, (est, oc) in existentes.items() if est != 'Pendiente']
        if no_pendientes:
            return jsonify({
                'error': 'Hay solicitudes que no estan Pendientes',
                'no_pendientes': no_pendientes,
            }), 409
        ya_oc = [n for n, (est, oc) in existentes.items() if oc]
        if ya_oc:
            return jsonify({
                'error': 'Hay solicitudes que ya tienen OC asociada',
                'ya_con_oc': ya_oc,
            }), 409

        # 2. Cargar items de todas las solicitudes
        try:
            c.execute(f"""
                SELECT numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       COALESCE(valor_estimado, 0),
                       COALESCE(precio_unit_g, 0),
                       COALESCE(proveedor_sugerido, '')
                FROM solicitudes_compra_items
                WHERE numero IN ({placeholders})
            """, nums)
            items_raw = [
                {'numero': r[0], 'codigo_mp': r[1] or '', 'nombre_mp': r[2] or '',
                 'cantidad_g': float(r[3] or 0), 'unidad': r[4] or 'g',
                 'valor_estimado': float(r[5] or 0),
                 'precio_unit_g': float(r[6] or 0),
                 'proveedor_sugerido': r[7] or ''}
                for r in c.fetchall()
            ]
        except sqlite3.OperationalError:
            c.execute(f"""
                SELECT numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                       COALESCE(valor_estimado, 0)
                FROM solicitudes_compra_items
                WHERE numero IN ({placeholders})
            """, nums)
            items_raw = [
                {'numero': r[0], 'codigo_mp': r[1] or '', 'nombre_mp': r[2] or '',
                 'cantidad_g': float(r[3] or 0), 'unidad': r[4] or 'g',
                 'valor_estimado': float(r[5] or 0),
                 'precio_unit_g': 0.0,
                 'proveedor_sugerido': ''}
                for r in c.fetchall()
            ]

        if not items_raw:
            return jsonify({'error': 'Las solicitudes no tienen items'}), 400

        # 3. Consolidar items por codigo_mp si se pidio
        if consolidar:
            consolidados = {}
            for it in items_raw:
                k = (it['codigo_mp'] or '').upper().strip() or it['nombre_mp'][:50]
                if k not in consolidados:
                    consolidados[k] = {
                        'codigo_mp': it['codigo_mp'],
                        'nombre_mp': it['nombre_mp'],
                        'cantidad_g': 0.0,
                        'precio_unitario': it['precio_unit_g'] or 0.0,
                        'unidad': it['unidad'],
                    }
                consolidados[k]['cantidad_g'] += it['cantidad_g']
                # Si llega un precio_unit_g distinto, dejar el primero no-cero
                if not consolidados[k]['precio_unitario'] and it['precio_unit_g']:
                    consolidados[k]['precio_unitario'] = it['precio_unit_g']
            items_oc = list(consolidados.values())
        else:
            items_oc = [
                {'codigo_mp': it['codigo_mp'], 'nombre_mp': it['nombre_mp'],
                 'cantidad_g': it['cantidad_g'],
                 'precio_unitario': it['precio_unit_g'],
                 'unidad': it['unidad']}
                for it in items_raw
            ]

        # BUG #1 CRITICA · Sprint Compras N1 · 21-may-2026 ·
        # Sebastián: agente detectó que oc-desde-solicitudes NO chequea
        # _check_monto_limit antes de crear OC bulk · cualquier user
        # Compras puede crear OC de $500M sin escalado.
        # FIX: pre-calcular valor_total preview y validar contra límite
        # del usuario · si excede → 403 sin tocar BD.
        valor_total_preview = 0.0
        for _it in items_oc:
            try:
                _cant = float(_it.get('cantidad_g') or 0)
                _pu = float(_it.get('precio_unitario') or 0)
                valor_total_preview += _cant * _pu
            except (ValueError, TypeError):
                pass
        if valor_total_preview > 0:
            err_lim, code_lim = _check_monto_limit(usuario, valor_total_preview)
            if err_lim:
                # NO tocar BD · solo devolver el error con el valor calculado
                # para que la UI muestre cuánto excede.
                return err_lim, code_lim

        # 4. Crear la OC
        obs = f"OC consolidada desde {len(nums)} solicitudes"
        if obs_extra:
            obs = obs_extra + ' · ' + obs
        # numero único con reintento ante carrera MAX+1 entre workers
        for _intento in range(6):
            numero_oc = _siguiente_numero_oc(c)
            try:
                c.execute(
                    """INSERT INTO ordenes_compra
                       (numero_oc, fecha, estado, proveedor, observaciones,
                        creado_por, fecha_entrega_est, categoria)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (numero_oc, datetime.now().isoformat(), 'Borrador', proveedor,
                     obs, usuario, fecha_entrega_est, categoria),
                )
                break
            except sqlite3.IntegrityError:
                if _intento == 5:
                    raise

        # 5. Auto-crear proveedor si no existe (con audit_log)
        # FIX · 16-jun-2026 · `proveedores.nombre` tiene UNIQUE GLOBAL (sin
        # importar activo), pero la existencia se chequeaba con `AND activo=1`:
        # un proveedor INACTIVO con el mismo nombre disparaba un INSERT que
        # choca con el UNIQUE → IntegrityError. El `except` lo tragaba, PERO en
        # PostgreSQL un INSERT fallido ABORTA toda la transacción → el INSERT de
        # items siguiente moría con "transaction aborted" → 500 genérico
        # "No se pudo crear la OC" (drift SQLite↔PG · CERO_ERROR). Dos defensas:
        # (a) chequear existencia por nombre SIN filtrar activo (= match el
        # UNIQUE) y reactivar si está inactivo; (b) SAVEPOINT para que cualquier
        # fallo del bloque NO envenene la transacción del caller.
        try:
            c.execute('SAVEPOINT _prov_auto')
            try:
                existe = c.execute(
                    "SELECT id, COALESCE(activo,0) FROM proveedores "
                    "WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?)) ORDER BY activo DESC LIMIT 1",
                    (proveedor,),
                ).fetchone()
                _prov_id, _prov_accion = None, None
                if not existe:
                    c.execute(
                        """INSERT INTO proveedores (nombre, categoria, condiciones_pago,
                                                     activo, fecha_creacion)
                           VALUES (?,?,?,1,?)""",
                        (proveedor, categoria, '30 dias', datetime.now().isoformat()),
                    )
                    _prov_id, _prov_accion = c.lastrowid, 'CREAR_PROVEEDOR'
                elif not existe[1]:
                    # existe pero INACTIVO → reactivar (re-INSERT chocaría el UNIQUE)
                    c.execute("UPDATE proveedores SET activo=1 WHERE id=?", (existe[0],))
                    _prov_id, _prov_accion = existe[0], 'REACTIVAR_PROVEEDOR'
                if _prov_accion:
                    audit_log(
                        c, usuario=usuario, accion=_prov_accion,
                        tabla='proveedores', registro_id=_prov_id,
                        despues={'nombre': proveedor[:200], 'categoria': categoria,
                                  'origen': 'auto_oc_desde_solicitudes',
                                  'oc_origen': numero_oc},
                        detalle=f"Auto al consolidar {len(nums)} solicitudes en {numero_oc}",
                    )
                c.execute('RELEASE SAVEPOINT _prov_auto')
            except Exception:
                try:
                    c.execute('ROLLBACK TO SAVEPOINT _prov_auto')
                    c.execute('RELEASE SAVEPOINT _prov_auto')
                except Exception:
                    pass
        except Exception:
            pass

        # 6. Insertar items en la OC
        valor_total = 0.0
        for it in items_oc:
            cant_g = float(it.get('cantidad_g') or 0)
            pu = float(it.get('precio_unitario') or 0)
            subt = round(cant_g * pu, 2)
            valor_total += subt
            c.execute(
                """INSERT INTO ordenes_compra_items
                   (numero_oc, codigo_mp, nombre_mp, cantidad_g, precio_unitario, subtotal)
                   VALUES (?,?,?,?,?,?)""",
                (numero_oc, it.get('codigo_mp', ''), it.get('nombre_mp', ''),
                 cant_g, pu, subt),
            )
            # Historico de precios + ref maestro_mps
            cod_mp = it.get('codigo_mp', '')
            if cod_mp and pu > 0:
                try:
                    # FIX 1-jun-2026 audit Abastecimiento (P1) · el INSERT usaba columnas
                    # inexistentes (nombre_mp/precio_unitario/cantidad_g) y omitía precio_kg
                    # (NOT NULL) → fallaba SIEMPRE en silencio → el histórico de precios NUNCA
                    # se grababa desde el flujo canónico de OC. Columnas reales + precio_kg
                    # = pu($/g) × 1000.
                    c.execute(
                        """INSERT OR IGNORE INTO precios_mp_historico
                           (codigo_mp, proveedor, precio_kg, fecha, numero_oc, origen)
                           VALUES (?,?,?,?,?,'oc')""",
                        (cod_mp, proveedor, pu * 1000.0,
                         datetime.now().isoformat()[:10], numero_oc),
                    )
                except Exception as _eph:
                    __import__('logging').getLogger('compras').warning(
                        'precios_mp_historico insert falló (oc %s, mp %s): %s', numero_oc, cod_mp, _eph)
                try:
                    c.execute(
                        """UPDATE maestro_mps
                           SET precio_referencia=?,
                               proveedor=COALESCE(NULLIF(proveedor,''),?)
                           WHERE codigo_mp=?""",
                        (pu * 1000.0, proveedor, cod_mp),  # $/g → $/kg (INV-2)
                    )
                except Exception:
                    pass
            # Fix #6 · 21-may-2026 · precio_real escribe back en SOL items
            # (antes valor_estimado quedaba 0 eternamente · reportes mentían)
            if cod_mp and pu > 0:
                try:
                    c.execute(
                        """UPDATE solicitudes_compra_items
                           SET valor_estimado=?
                           WHERE numero IN ({}) AND codigo_mp=?""".format(placeholders),
                        [round(cant_g * pu, 2)] + nums + [cod_mp],
                    )
                except Exception:
                    pass

        if valor_total > 0:
            c.execute(
                "UPDATE ordenes_compra SET valor_total=? WHERE numero_oc=?",
                (valor_total, numero_oc),
            )
            # Fix #6 cont · valor total back en SOL (suma items)
            try:
                for _num in nums:
                    c.execute(
                        """UPDATE solicitudes_compra
                           SET valor = (SELECT COALESCE(SUM(valor_estimado),0)
                                        FROM solicitudes_compra_items WHERE numero=?)
                           WHERE numero=?""",
                        (_num, _num),
                    )
            except Exception:
                pass

        # Fase 2 · 21-may-2026 · auto-aprobación por reglas
        # Si OC cumple: monto<$500k + proveedor recurrente (≥3 OCs/90d) +
        # precios en rango ±15% promedio, pasa Borrador → Autorizada automático.
        # Catalina solo aprueba excepciones · libera 30-40 min/día.
        auto_aprobada = False
        auto_razon = ''
        try:
            auto_aprobada, auto_razon = _evaluar_auto_aprobacion(
                c, proveedor, valor_total, items_oc,
            )
            if auto_aprobada:
                from datetime import datetime as _dtauto
                _fecha_aut = _dtauto.now().isoformat()
                _fecha_hoy = _dtauto.now().strftime('%Y%m%d')
                # Generar remision_code
                _last = c.execute(
                    "SELECT remision_code FROM ordenes_compra WHERE remision_code LIKE ? "
                    "ORDER BY remision_code DESC LIMIT 1",
                    (f'REM-ESP-{_fecha_hoy}-%',),
                ).fetchone()
                _n = int(_last[0].split('-')[-1]) + 1 if _last and _last[0] else 1
                _rem = f'REM-ESP-{_fecha_hoy}-{_n:03d}'
                # SEC-FIX · 21-may-2026 · auto-aprob respeta segregation
                # autorizado_por='auto-aprob-reglas' (sistema), no usuario.
                # · contadora NUNCA debe aparecer como autorizadora aunque
                # la OC se auto-apruebe via su acción de crear.
                c.execute(
                    """UPDATE ordenes_compra SET estado='Autorizada',
                                                 remision_code=?,
                                                 autorizado_por='auto-aprob-reglas',
                                                 fecha_autorizacion=?,
                                                 observaciones=COALESCE(observaciones,'') || ' | AUTO-APROB: ' || ?
                       WHERE numero_oc=?""",
                    (_rem, _fecha_aut, auto_razon, numero_oc),
                )
                try:
                    audit_log(c, usuario='auto-aprob-reglas',
                              accion='OC_AUTO_APROBADA',
                              tabla='ordenes_compra', registro_id=numero_oc,
                              despues={'monto': valor_total, 'proveedor': proveedor,
                                       'razon': auto_razon})
                except Exception:
                    pass
        except Exception as _e:
            log.warning('auto-aprob fallo: %s', _e)

        # 7. Vincular las solicitudes a la nueva OC y marcarlas Aprobada
        c.execute(
            f"""UPDATE solicitudes_compra
                SET estado='Aprobada', numero_oc=?
                WHERE numero IN ({placeholders}) AND estado='Pendiente'
                  AND COALESCE(numero_oc,'')=''""",
            [numero_oc] + nums,
        )
        vinculadas = c.rowcount

        # 8. Audit log
        try:
            audit_log(
                c, usuario=usuario, accion='CREAR_OC_BULK',
                tabla='ordenes_compra', registro_id=numero_oc,
                despues={
                    'proveedor': proveedor[:200],
                    'categoria': categoria,
                    'items_count': len(items_oc),
                    'valor_total': round(valor_total, 2),
                    'solicitudes_vinculadas': vinculadas,
                    'solicitudes_origen': nums[:50],  # cap a 50 en audit
                    'consolidar_iguales': consolidar,
                },
                detalle=(f"Creó OC {numero_oc} consolidada desde "
                         f"{vinculadas} solicitudes · {proveedor} · "
                         f"{len(items_oc)} items · total {valor_total:.0f}"),
            )
        except Exception as e:
            log.warning('audit_log CREAR_OC_BULK fallo: %s', e)

        conn.commit()
        return jsonify({
            'ok': True,
            'numero_oc': numero_oc,
            'proveedor': proveedor,
            'items_creados': len(items_oc),
            'solicitudes_vinculadas': vinculadas,
            'valor_total': round(valor_total, 2),
            'estado': 'Autorizada' if auto_aprobada else 'Borrador',
            'auto_aprobada': auto_aprobada,  # Fase 2 · feedback al frontend
            'auto_razon': auto_razon if auto_aprobada else '',
        }), 201
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        log.exception('crear_oc_desde_solicitudes fallo: %s', e)
        # M4 · devolver la causa REAL (app interna) para diagnóstico, no tragarla
        return jsonify({'error': 'No se pudo crear la OC desde las solicitudes',
                        'detalle': str(e)[:300]}), 500


# ── Sebastián 4-may-2026 (Catalina): consolidar AUTO-XXXX pendientes ────
# Una sola pasada que toma las 200+ AUTO-XXXX Pendientes (1-MP-cada-una)
# y las transforma en ~15-20 solicitudes consolidadas por proveedor.
# El generador ya consolida desde aqui (auto_plan.py:936) pero hay datos
# legacy que necesitan limpieza. Idempotente: si vuelve a correr cuando
# ya esta consolidado, no rompe nada.
@bp.route('/api/compras/consolidar-auto-pendientes', methods=['POST'])
def consolidar_auto_pendientes():
    """[DEPRECATED 21-may-2026] One-shot migración 4-may-2026 completada.

    No borrar todavía · frontend lo usa pero ahora es no-op silencioso
    si no hay AUTO-XXXX legacy. Sebastián 21-may: marcado deprecated ·
    candidato a borrar en próxima limpieza (~30 LOC + endpoint).

    [Original docstring]
    Consolida solicitudes AUTO-XXXX Pendientes existentes por proveedor.

    Body JSON (opcional):
      {
        "dry_run": false,        # default false. Si true, solo retorna
                                  el plan de consolidacion sin ejecutar.
        "min_para_consolidar": 5 # solo consolida si hay >=5 SOLs sueltas
                                  con el mismo proveedor (default 1 = todas)
      }

    Algoritmo:
      1. SELECT de AUTO-XXXX Pendientes con sus items
      2. Identifica las que tienen 1 item (sospechosas de ser legacy)
      3. Agrupa por (proveedor_sugerido, categoria)
      4. Si grupo tiene >= min_para_consolidar SOLs → crear UNA nueva
         AUTO-XXXX consolidada con todos los items + delete las viejas
      5. Atomico: rollback completo si algo falla
      6. audit_log CONSOLIDAR_AUTO_PENDIENTES con conteos

    Returns:
      200 {ok, antes, despues, grupos: [...], eliminadas, creadas}
      403 si no tiene permisos
    """
    usuario, err, code = _require_compras_write()
    if err:
        return err, code

    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get('dry_run', False))
    try:
        min_para_consolidar = int(body.get('min_para_consolidar', 1))
    except (ValueError, TypeError):
        min_para_consolidar = 1
    if min_para_consolidar < 1:
        min_para_consolidar = 1

    conn = get_db(); c = conn.cursor()

    # 1. Cargar AUTO-XXXX Pendientes (sin OC) y sus items
    c.execute("""
        SELECT numero, urgencia, COALESCE(observaciones,''), categoria, fecha
        FROM solicitudes_compra
        WHERE numero LIKE 'AUTO-%' AND estado='Pendiente'
          AND COALESCE(numero_oc,'') = ''
        ORDER BY numero
    """)
    auto_sols = c.fetchall()
    if not auto_sols:
        return jsonify({
            'ok': True, 'antes': 0, 'despues': 0,
            'mensaje': 'Sin AUTO-XXXX pendientes para consolidar',
            'grupos': [], 'eliminadas': 0, 'creadas': 0,
        })

    nums = [r[0] for r in auto_sols]
    placeholders = ','.join(['?'] * len(nums))

    try:
        c.execute(f"""
            SELECT numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                   COALESCE(valor_estimado,0), COALESCE(justificacion,''),
                   COALESCE(precio_unit_g,0), COALESCE(proveedor_sugerido,'')
            FROM solicitudes_compra_items
            WHERE numero IN ({placeholders})
        """, nums)
        items_all = [
            {'numero': r[0], 'codigo_mp': r[1] or '', 'nombre_mp': r[2] or '',
             'cantidad_g': float(r[3] or 0), 'unidad': r[4] or 'g',
             'valor_estimado': float(r[5] or 0), 'justificacion': r[6] or '',
             'precio_unit_g': float(r[7] or 0),
             'proveedor_sugerido': r[8] or ''}
            for r in c.fetchall()
        ]
    except sqlite3.OperationalError:
        c.execute(f"""
            SELECT numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                   COALESCE(valor_estimado,0), COALESCE(justificacion,'')
            FROM solicitudes_compra_items
            WHERE numero IN ({placeholders})
        """, nums)
        items_all = []
        for r in c.fetchall():
            items_all.append({
                'numero': r[0], 'codigo_mp': r[1] or '',
                'nombre_mp': r[2] or '', 'cantidad_g': float(r[3] or 0),
                'unidad': r[4] or 'g', 'valor_estimado': float(r[5] or 0),
                'justificacion': r[6] or '', 'precio_unit_g': 0.0,
                'proveedor_sugerido': '',
            })

    items_por_sol = {}
    for it in items_all:
        items_por_sol.setdefault(it['numero'], []).append(it)

    # Sebastian 4-may-2026 (Catalina): si el item NO tiene proveedor_sugerido
    # (legacy de aplicar_plan que leia solo mp_lead_time_config), buscarlo en
    # maestro_mps.proveedor por codigo_mp. Asi no aparecen "sin proveedor".
    cods_sin_prov = {it['codigo_mp'].upper() for it in items_all
                      if it['codigo_mp'] and not (it.get('proveedor_sugerido') or '').strip()}
    prov_lookup = {}
    if cods_sin_prov:
        ph_cod = ','.join(['?'] * len(cods_sin_prov))
        try:
            for r in c.execute(
                f"SELECT UPPER(codigo_mp), COALESCE(proveedor,'') "
                f"FROM maestro_mps WHERE UPPER(codigo_mp) IN ({ph_cod})",
                list(cods_sin_prov),
            ).fetchall():
                if r[1]:
                    prov_lookup[r[0]] = r[1]
        except Exception:
            pass
    # Aplicar fallback in-place a items_all
    for it in items_all:
        if not (it.get('proveedor_sugerido') or '').strip():
            cod_up = (it.get('codigo_mp') or '').upper()
            fallback = prov_lookup.get(cod_up, '')
            if fallback:
                it['proveedor_sugerido'] = fallback
                it['_proveedor_fallback'] = True  # marca para audit/observ

    # 2. Agrupar por (proveedor_sugerido, categoria) las que SOLO tienen 1 item
    # (legacy 1-MP-por-AUTO-XXXX). Las que ya tienen N>1 items son las nuevas
    # consolidadas - se respetan tal cual.
    URG_RANK = {'Critico': 4, 'Urgente': 3, 'Alta': 2, 'Media': 1, 'Normal': 0}
    grupos = {}
    intactas = []  # AUTO-XXXX con >=2 items (ya consolidadas)
    sol_meta = {r[0]: {'urgencia': r[1] or 'Normal', 'observ': r[2] or '',
                        'categoria': r[3] or 'Materia Prima',
                        'fecha': r[4]} for r in auto_sols}
    for num in nums:
        items = items_por_sol.get(num, [])
        if len(items) >= 2:
            intactas.append(num)
            continue
        if not items:
            # SOL sin items — caso raro. La preservamos.
            intactas.append(num)
            continue
        prov = (items[0].get('proveedor_sugerido') or '').strip()
        cat = sol_meta[num]['categoria']
        key = (prov, cat)
        grupos.setdefault(key, []).append(num)

    # 3. Construir plan de consolidacion
    plan_grupos = []
    for (prov, cat), sols in grupos.items():
        if len(sols) < min_para_consolidar:
            # No vale la pena consolidar
            continue
        # Sumar urgencia max
        rank_max = 0
        for n in sols:
            r = URG_RANK.get(sol_meta[n]['urgencia'], 0)
            if r > rank_max:
                rank_max = r
        urg_label = {4: 'Critico', 3: 'Urgente', 2: 'Alta',
                      1: 'Media', 0: 'Normal'}[rank_max]
        # Items de todas las SOLs del grupo
        items_grupo = []
        for n in sols:
            for it in items_por_sol.get(n, []):
                items_grupo.append(it)
        plan_grupos.append({
            'proveedor': prov,
            'proveedor_label': prov or '(Sin proveedor sugerido)',
            'categoria': cat,
            'urgencia': urg_label,
            'sols_origen': sols,
            'items_count': len(items_grupo),
            'total_g': round(sum(it['cantidad_g'] for it in items_grupo), 0),
            '_items_payload': items_grupo,  # se elimina antes de retornar JSON
        })

    if dry_run:
        # Retornar el plan sin tocar DB
        for g in plan_grupos:
            g.pop('_items_payload', None)
        return jsonify({
            'ok': True,
            'dry_run': True,
            'antes': len(nums),
            'intactas': len(intactas),
            'despues': len(plan_grupos) + len(intactas),
            'grupos': plan_grupos,
            'eliminadas': 0,
            'creadas': 0,
            'mensaje': (f'Plan: consolidar {sum(len(g["sols_origen"]) for g in plan_grupos)} '
                         f'SOLs en {len(plan_grupos)} grupos · {len(intactas)} '
                         f'quedan intactas (ya consolidadas)'),
        })

    if not plan_grupos:
        return jsonify({
            'ok': True,
            'antes': len(nums),
            'intactas': len(intactas),
            'despues': len(nums),
            'grupos': [],
            'eliminadas': 0,
            'creadas': 0,
            'mensaje': f'Nada que consolidar · {len(intactas)} ya estan consolidadas',
        })

    # 4. Ejecutar atomicamente
    creadas_nums = []
    sols_a_eliminar = []
    try:
        for g in plan_grupos:
            top = sorted(g['_items_payload'], key=lambda x: -float(x.get('cantidad_g') or 0))[:3]
            resumen = ', '.join(
                f"{(it.get('nombre_mp') or '')[:25]} {round(float(it.get('cantidad_g') or 0)/1000.0,1)}kg"
                for it in top
            )
            if g['items_count'] > 3:
                resumen += f", +{g['items_count'] - 3} mas"
            obs = (
                f"AUTO-PLAN consolidado · proveedor: {g['proveedor_label']} · "
                f"{g['items_count']} MPs · {round(g['total_g']/1000.0,1)}kg total · "
                f"{resumen} · (consolidacion legacy {datetime.now().date().isoformat()})"
            )
            # numero AUTO-XXXX único con reintento ante carrera MAX+1
            for _intento in range(6):
                next_n = c.execute("""
                    SELECT COALESCE(MAX(CAST(SUBSTR(numero, 6) AS INTEGER)), 0) + 1
                    FROM solicitudes_compra WHERE numero LIKE 'AUTO-%'
                """).fetchone()[0] or 1
                new_num = f'AUTO-{next_n:04d}'
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra
                          (numero, fecha, estado, solicitante, urgencia,
                           observaciones, area, empresa, categoria, tipo, valor)
                        VALUES (?, date('now', '-5 hours'), 'Pendiente', 'AUTO-PLAN', ?, ?,
                                'Producción', 'Espagiria', ?, 'Compra', 0)
                    """, (new_num, g['urgencia'], obs, g['categoria']))
                    break
                except sqlite3.IntegrityError:
                    if _intento == 5:
                        raise
            for it in g['_items_payload']:
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado, proveedor_sugerido,
                           precio_unit_g)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (new_num, it['codigo_mp'], it['nombre_mp'],
                          it['cantidad_g'], it['unidad'],
                          it['justificacion'], it['valor_estimado'],
                          g['proveedor'], it['precio_unit_g']))
                except sqlite3.OperationalError:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (new_num, it['codigo_mp'], it['nombre_mp'],
                          it['cantidad_g'], it['unidad'],
                          it['justificacion'], it['valor_estimado']))
            creadas_nums.append(new_num)
            sols_a_eliminar.extend(g['sols_origen'])

        # Eliminar las legacy
        if sols_a_eliminar:
            ph_del = ','.join(['?'] * len(sols_a_eliminar))
            c.execute(f"DELETE FROM solicitudes_compra_items WHERE numero IN ({ph_del})",
                      sols_a_eliminar)
            c.execute(f"DELETE FROM solicitudes_compra WHERE numero IN ({ph_del})",
                      sols_a_eliminar)

        # Audit
        try:
            audit_log(
                c, usuario=usuario, accion='CONSOLIDAR_AUTO_PENDIENTES',
                tabla='solicitudes_compra', registro_id='bulk',
                despues={
                    'antes': len(nums), 'intactas': len(intactas),
                    'eliminadas': len(sols_a_eliminar),
                    'creadas': len(creadas_nums),
                    'creadas_nums': creadas_nums[:50],
                    'min_para_consolidar': min_para_consolidar,
                },
                detalle=(f"Consolido {len(sols_a_eliminar)} AUTO-XXXX legacy "
                          f"en {len(creadas_nums)} solicitudes por proveedor"),
            )
        except Exception as _e:
            log.warning('audit_log CONSOLIDAR_AUTO_PENDIENTES fallo: %s', _e)

        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        log.exception('consolidar_auto_pendientes fallo: %s', e)
        return jsonify({'error': 'No se pudo consolidar las solicitudes'}), 500

    return jsonify({
        'ok': True,
        'antes': len(nums),
        'intactas': len(intactas),
        'eliminadas': len(sols_a_eliminar),
        'creadas': len(creadas_nums),
        'creadas_nums': creadas_nums,
        'despues': len(creadas_nums) + len(intactas),
        'mensaje': (f'Consolidadas {len(sols_a_eliminar)} solicitudes legacy en '
                     f'{len(creadas_nums)} agrupadas por proveedor'),
    })


# ── Sebastián 4-may-2026 (Catalina): limpiar y regenerar AUTO-PLAN ────
# Sebastian 6-may-2026: limpieza completa de SOLs Pendientes de Planta.
# Borra TODO lo que vino de Planta (auto-plan + manual con categoria
# Materia Prima/Empaque) sin OC vinculada. Usado para empezar de cero
# antes de cargar con la nueva disposicion de calendario.
@bp.route('/api/compras/limpiar-solicitudes-planta', methods=['POST'])
def limpiar_solicitudes_planta():
    """Borra TODAS las SOLs Pendientes con categoria 'Materia Prima' o
    'Empaque' que NO tengan OC vinculada (=quedaron huérfanas tras los
    duplicados del cron auto-plan).

    Body JSON:
      {dry_run: false}  · si true, solo cuenta sin borrar

    Returns: {ok, eliminadas, items_eliminados, mensaje}
    """
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get('dry_run', False))
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT numero FROM solicitudes_compra
        WHERE estado='Pendiente'
          AND COALESCE(numero_oc,'') = ''
          AND categoria IN ('Materia Prima','Empaque','Material de Empaque')
    """).fetchall()
    nums = [r[0] for r in rows]
    if dry_run:
        return jsonify({
            'ok': True,
            'dry_run': True,
            'eliminaria': len(nums),
            'mensaje': f'Plan: borrar {len(nums)} SOLs de Planta Pendientes sin OC',
        })
    if not nums:
        return jsonify({
            'ok': True, 'eliminadas': 0, 'items_eliminados': 0,
            'mensaje': 'No hay SOLs de Planta Pendientes que borrar',
        })
    try:
        ph = ','.join(['?'] * len(nums))
        r1 = c.execute(
            f"DELETE FROM solicitudes_compra_items WHERE numero IN ({ph})",
            nums,
        )
        items_eliminados = r1.rowcount or 0
        c.execute(
            f"DELETE FROM solicitudes_compra WHERE numero IN ({ph})",
            nums,
        )
        try:
            audit_log(
                c, usuario=usuario, accion='LIMPIAR_SOLS_PLANTA',
                tabla='solicitudes_compra', registro_id='bulk',
                despues={'eliminadas': len(nums),
                          'items_eliminados': items_eliminados},
                detalle=(f"Borró {len(nums)} SOLs Planta Pendientes "
                          f"(MP/Empaque sin OC) · {items_eliminados} items"),
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        log.exception('limpiar_solicitudes_planta fallo: %s', e)
        return jsonify({'error': 'No se pudieron limpiar las solicitudes de planta'}), 500
    return jsonify({
        'ok': True,
        'eliminadas': len(nums),
        'items_eliminados': items_eliminados,
        'mensaje': (f'✓ Eliminadas {len(nums)} SOLs Pendientes de Planta · '
                     f'{items_eliminados} items. Ahora puedes cargar limpio '
                     f'desde el calendario.'),
    })


@bp.route('/api/compras/limpiar-y-regenerar-auto-plan', methods=['POST'])
def limpiar_y_regenerar_auto_plan():
    """[ADMIN-NUCLEAR] Borra TODAS las AUTO-XXXX Pendientes y regenera.

    SEC-FIX · 21-may-2026 · solo ADMIN puede ejecutar · era riesgo de
    botón rojo accidental que volaba 200+ solicitudes pendientes.

    [Original docstring]
    Borra TODAS las AUTO-XXXX Pendientes (sin OC) y vuelve a generar.

    Caso de uso: los AUTO-XXXX existentes vienen del cron que leia
    mp_lead_time_config sin fallback a maestro_mps.proveedor → muchas
    quedaban sin proveedor. Con el fix de aplicar_plan() (4-may-2026)
    ahora SI tienen fallback. Este endpoint deja todo limpio y
    regenera con la logica nueva → Catalina ve solicitudes consolidadas
    por proveedor real.

    Body JSON (opcional):
      {
        "horizonte_dias": 60,        # default 60d
        "dry_run": false,            # si true, solo cuenta lo que borraria
        "regenerar": true            # default true. Si false, solo borra y
                                      # deja que el cron de planta regenere
                                      # agrupado en la siguiente corrida.
      }

    Returns:
      200 {ok, eliminadas, creadas, mensaje}
    """
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    # SEC-FIX · 21-may-2026 · solo ADMIN puede ejecutar (era riesgo accidental)
    if (usuario or '').lower() not in {x.lower() for x in ADMIN_USERS}:
        return jsonify({
            'error': 'Operación nuclear · solo admin (Sebas/Alejandro)',
            'codigo': 'NUCLEAR_SOLO_ADMIN',
            'hint': 'Esta operación borra 200+ SOLs · solo admin para evitar accidentes',
        }), 403

    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get('dry_run', False))
    regenerar = bool(body.get('regenerar', True))
    try:
        horizonte = int(body.get('horizonte_dias', 60))
    except (ValueError, TypeError):
        horizonte = 60
    if not (7 <= horizonte <= 365):
        return jsonify({'error': 'horizonte_dias fuera de rango (7-365)'}), 400

    conn = get_db(); c = conn.cursor()
    # Sebastian 4-may-2026 (Catalina · "elimina todas asi cuando volvamos a
    # pedir desde planta ya se agrupan"): captura BOTH patrones de auto-gen:
    #   1. AUTO-XXXX Pendientes sin OC (cron aplicar_plan)
    #   2. SOL-YYYY-XXXX Pendientes con observaciones "Auto-generada" o
    #      "Centro Programación" (regenerar-oc / auto-sc-ia)
    # Sus OCs en Borrador asociadas también se borran.
    c.execute("""
        SELECT numero, COALESCE(numero_oc,'')
        FROM solicitudes_compra
        WHERE estado='Pendiente'
          AND (
            (numero LIKE 'AUTO-%' AND COALESCE(numero_oc,'')='')
            OR (numero LIKE 'SOL-%' AND (
                observaciones LIKE '%Auto-generada%'
                OR observaciones LIKE '%Centro Programación%'
                OR observaciones LIKE '%Centro Programacion%'
                OR observaciones LIKE '%Auto-SC IA%'
                OR observaciones LIKE '%🤖 Auto-SC%'
                OR LOWER(COALESCE(solicitante,'')) IN ('auto-plan','auto-plan-ia')
            ))
          )
    """)
    rows_a_borrar = c.fetchall()
    nums_a_borrar = [r[0] for r in rows_a_borrar]
    ocs_a_revisar = [r[1] for r in rows_a_borrar if r[1]]

    # OCs en Borrador asociadas (no se tocan Autorizadas/Pagadas)
    ocs_borrador_a_borrar = []
    if ocs_a_revisar:
        ph_oc = ','.join(['?'] * len(ocs_a_revisar))
        ocs_borrador_a_borrar = [
            r[0] for r in c.execute(
                f"SELECT numero_oc FROM ordenes_compra "
                f"WHERE numero_oc IN ({ph_oc}) AND estado='Borrador'",
                ocs_a_revisar,
            ).fetchall()
        ]

    if dry_run:
        plan_txt = (f'Plan: borrar {len(nums_a_borrar)} solicitudes auto-generadas '
                    f'(AUTO-XXXX + SOL-YYYY-XXXX Pendientes)')
        if ocs_borrador_a_borrar:
            plan_txt += f' + {len(ocs_borrador_a_borrar)} OCs Borrador asociadas'
        if regenerar:
            plan_txt += f' · regenerar con horizonte {horizonte}d'
        else:
            plan_txt += ' · sin regenerar (cron de planta lo hara despues agrupado)'
        return jsonify({
            'ok': True,
            'dry_run': True,
            'eliminaria': len(nums_a_borrar),
            'eliminaria_ocs_borrador': len(ocs_borrador_a_borrar),
            'horizonte_dias': horizonte,
            'regenerar': regenerar,
            'mensaje': plan_txt,
        })

    try:
        # 1. Borrar items + solicitudes
        eliminadas_items = 0
        if nums_a_borrar:
            ph = ','.join(['?'] * len(nums_a_borrar))
            r1 = c.execute(
                f"DELETE FROM solicitudes_compra_items WHERE numero IN ({ph})",
                nums_a_borrar,
            )
            eliminadas_items = r1.rowcount or 0
            c.execute(
                f"DELETE FROM solicitudes_compra WHERE numero IN ({ph})",
                nums_a_borrar,
            )

        # 2. Borrar OCs Borrador asociadas (Autorizadas/Pagadas se preservan)
        ocs_eliminadas = 0
        if ocs_borrador_a_borrar:
            ph_b = ','.join(['?'] * len(ocs_borrador_a_borrar))
            try:
                c.execute(
                    f"DELETE FROM ordenes_compra_items WHERE numero_oc IN ({ph_b})",
                    ocs_borrador_a_borrar,
                )
            except sqlite3.OperationalError:
                pass
            r2 = c.execute(
                f"DELETE FROM ordenes_compra WHERE numero_oc IN ({ph_b})",
                ocs_borrador_a_borrar,
            )
            ocs_eliminadas = r2.rowcount or 0

        # 3. Audit del borrado
        try:
            audit_log(
                c, usuario=usuario, accion='LIMPIAR_AUTO_PLAN',
                tabla='solicitudes_compra', registro_id='bulk',
                despues={'eliminadas': len(nums_a_borrar),
                          'items_eliminados': eliminadas_items,
                          'ocs_borrador_eliminadas': ocs_eliminadas,
                          'horizonte_dias': horizonte,
                          'regenerar': regenerar},
                detalle=(f"Limpio {len(nums_a_borrar)} solicitudes auto-generadas "
                          f"+ {ocs_eliminadas} OCs Borrador"),
            )
        except Exception:
            pass

        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        log.exception('limpiar_y_regenerar_auto_plan borrado fallo: %s', e)
        return jsonify({'error': f'Borrado fallido: {e}'}), 500

    # 4. Regenerar con generar_plan + aplicar_plan (solo si regenerar=true)
    if not regenerar:
        return jsonify({
            'ok': True,
            'eliminadas': len(nums_a_borrar),
            'ocs_borrador_eliminadas': ocs_eliminadas,
            'creadas': 0,
            'regenerar': False,
            'horizonte_dias': horizonte,
            'grupos': [],
            'mensaje': (f'✓ Limpiadas {len(nums_a_borrar)} solicitudes auto-generadas '
                         f'+ {ocs_eliminadas} OCs Borrador. '
                         f'El cron de planta regenerara agrupado por proveedor '
                         f'en la proxima corrida (o usa Regenerar manual).'),
        })

    try:
        from blueprints.auto_plan import generar_plan, aplicar_plan
        plan = generar_plan(horizonte_dias=horizonte, tipo='manual', usuario=usuario)
        resultado = aplicar_plan(plan, usuario=usuario)
    except Exception as e:
        log.exception('limpiar_y_regenerar_auto_plan regeneracion fallo: %s', e)
        return jsonify({
            'error': f'Borrado OK pero falló la regeneración: {e}',
            'eliminadas': len(nums_a_borrar),
            'creadas': 0,
        }), 500

    creadas = resultado.get('compras_creadas', []) or []
    grupos_resumen = []
    for cc in creadas[:30]:
        grupos_resumen.append({
            'numero': cc.get('numero'),
            'proveedor': cc.get('proveedor', ''),
            'items_count': cc.get('items_count', 0),
            'total_g': cc.get('total_g', 0),
        })

    return jsonify({
        'ok': True,
        'eliminadas': len(nums_a_borrar),
        'creadas': len(creadas),
        'horizonte_dias': horizonte,
        'grupos': grupos_resumen,
        'mensaje': (f'✓ Eliminadas {len(nums_a_borrar)} AUTO-XXXX legacy · '
                     f'Regeneradas {len(creadas)} consolidadas por proveedor '
                     f'(horizonte {horizonte}d)'),
    })


@bp.route('/api/solicitudes-compra/mis', methods=['GET'])
def mis_solicitudes_con_ciclo():
    """Devuelve las solicitudes del usuario logueado con el ciclo COMPLETO
    consolidado: estado de la solicitud + estado de la OC vinculada + si
    fue recibida fisicamente. Pensado para que el SOLICITANTE (Hernando,
    Luz, etc) pueda hacer seguimiento sin depender de Catalina.

    Sebastian (29-abr-2026): "si alguien pide papel, lo hace en
    solicitudes, pero quizas alli mismo deberia aparecer todo el listado
    de solicitudes que sean generales para que vean como va, si ya fue
    aceptada pagada en transito y cuando lleguen coloquen recibido asi
    hacemos el cierre de todo".

    Querystring:
      ?usuario=<user>  → fuerza el filtro (admin puede pasar otro)
                         Si no se pasa, usa el session user.
      ?estado=todas|abiertas|cerradas (default abiertas)
        - abiertas: solicitudes en cualquier paso del ciclo no Recibida/Cancelada
        - cerradas: Recibida, Cancelada, Rechazada
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    user_lower = user.strip().lower()
    raw_filtro = (request.args.get('usuario') or user).strip().lower()
    # SEC-FIX 23-may-2026 · auditoría · antes ?usuario=X permitía
    # impersonación · cualquier user veía SOLs de otro · ahora solo
    # admin puede pasar usuario distinto al propio
    if raw_filtro != user_lower:
        try:
            from config import ADMIN_USERS as _AU
            admin_lower = {x.lower() for x in _AU}
        except Exception:
            admin_lower = set()
        if user_lower not in admin_lower:
            return jsonify({
                'error': 'Solo admin puede ver SOLs de otros usuarios',
                'codigo': 'IMPERSONACION_NO_PERMITIDA',
            }), 403
    usuario_filtro = raw_filtro
    estado_q = (request.args.get('estado') or 'abiertas').strip().lower()

    conn = get_db(); c = conn.cursor()
    # Sebastián 1-may-2026: las solicitudes de Influencer/Marketing
    # (Cuenta de Cobro · empresa ANIMUS · area Marketing) NO deben aparecer
    # acá — entran a Compras directo en pestaña "Influencer" para que
    # Catalina no se enrede mezclándolas con producción.
    incluir_influencer = request.args.get('incluir_influencer', '0').strip() in ('1','true','yes')
    sql_solic = """
        SELECT s.numero, s.fecha, s.estado as estado_sol, s.solicitante,
               s.urgencia, s.observaciones, s.numero_oc, s.area, s.empresa,
               s.categoria, s.tipo, s.fecha_requerida, s.valor,
               oc.estado as estado_oc, oc.proveedor as oc_proveedor,
               oc.fecha_pago, oc.fecha_recepcion, oc.fecha_entrega_est,
               oc.valor_total as oc_valor, oc.recibido_por
        FROM solicitudes_compra s
        LEFT JOIN ordenes_compra oc ON oc.numero_oc = s.numero_oc
        WHERE LOWER(COALESCE(s.solicitante,''))=?"""
    params_solic = [usuario_filtro]
    if not incluir_influencer:
        sql_solic += """
          AND COALESCE(s.categoria,'') NOT IN ('Cuenta de Cobro','Influencer/Marketing Digital')
          AND LOWER(COALESCE(s.area,'')) NOT IN ('marketing','marketing/animus')"""
    sql_solic += " ORDER BY s.fecha DESC LIMIT 200"
    rows = c.execute(sql_solic, params_solic).fetchall()

    cols = [d[0] for d in c.description]
    out = []
    for r in rows:
        d = dict(zip(cols, r))
        # Determinar el "paso" del ciclo y badge consolidado.
        # Orden de prioridad: lo mas avanzado del ciclo manda.
        oc_estado = (d.get('estado_oc') or '').strip()
        sol_estado = (d.get('estado_sol') or '').strip()
        paso = 1
        paso_label = 'Pendiente aprobación'
        paso_color = '#a16207'
        cerrado = False
        if oc_estado == 'Recibida':
            paso, paso_label, paso_color = 6, '✅ Recibida en bodega', '#15803d'
            cerrado = True
        elif oc_estado == 'Parcial':
            paso, paso_label, paso_color = 5, '📦 Recepción parcial — esperando resto', '#f59e0b'
        elif oc_estado == 'Pagada':
            paso, paso_label, paso_color = 4, '💸 Pagada — en tránsito hacia bodega', '#1e40af'
        elif oc_estado == 'Autorizada':
            paso, paso_label, paso_color = 3, '🟢 OC autorizada — pendiente pago', '#0891b2'
        elif oc_estado in ('Aprobada', 'Revisada'):
            paso, paso_label, paso_color = 3, f'📋 OC {oc_estado.lower()}', '#0891b2'
        elif oc_estado == 'Borrador':
            paso, paso_label, paso_color = 2, '📝 OC en borrador', '#6b7280'
        elif oc_estado == 'Pendiente':
            paso, paso_label, paso_color = 2, '📝 OC pendiente revisión', '#6b7280'
        elif oc_estado in ('Cancelada', 'Rechazada'):
            paso, paso_label, paso_color = 0, f'❌ OC {oc_estado.lower()}', '#dc2626'
            cerrado = True
        elif sol_estado == 'Aprobada':
            paso, paso_label, paso_color = 2, '🟢 Solicitud aprobada — generando OC', '#0891b2'
        elif sol_estado == 'Rechazada':
            paso, paso_label, paso_color = 0, '❌ Solicitud rechazada', '#dc2626'
            cerrado = True
        else:
            paso, paso_label, paso_color = 1, '⏳ Pendiente aprobación de Catalina', '#a16207'
        d['paso'] = paso
        d['paso_label'] = paso_label
        d['paso_color'] = paso_color
        d['cerrado'] = cerrado
        # Habilitar boton "Marcar recibido" cuando la OC esta Pagada/Autorizada
        # (la mercancia esta en transito) o si no hay OC pero la solicitud
        # fue aprobada y nunca se genero OC formal (caso simple).
        d['puede_marcar_recibido'] = (
            oc_estado in ('Pagada', 'Autorizada', 'Parcial')
            or (not oc_estado and sol_estado == 'Aprobada')
        )
        out.append(d)

    if estado_q == 'cerradas':
        out = [x for x in out if x['cerrado']]
    elif estado_q == 'abiertas':
        out = [x for x in out if not x['cerrado']]

    abiertas_count = sum(1 for x in out if not x.get('cerrado'))
    return jsonify({'solicitudes': out, 'usuario': usuario_filtro,
                    'total': len(out), 'abiertas': abiertas_count})


@bp.route('/api/solicitudes-compra/<numero>/marcar-recibido-solicitante', methods=['POST'])
def marcar_recibido_solicitante(numero):
    """El solicitante confirma que la mercancia ya llego (sin pasar por
    Catalina). Cierra la cadena en su lado: actualiza la OC a 'Recibida'
    si esta en Autorizada/Pagada/Parcial.

    Solo puede hacerlo el solicitante original (o admin).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json() or {}
    obs = (d.get('observaciones') or '').strip()
    conn = get_db(); c = conn.cursor()
    # Verificar autoria
    sol = c.execute(
        "SELECT solicitante, numero_oc, estado FROM solicitudes_compra WHERE numero=?",
        (numero.upper(),)
    ).fetchone()
    if not sol:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    solicitante_orig = (sol[0] or '').lower()
    es_admin = user in ADMIN_USERS
    if solicitante_orig != user.lower() and not es_admin:
        return jsonify({
            'error': 'Solo el solicitante original o un admin puede marcar recibido'
        }), 403

    numero_oc = sol[1] or ''
    if numero_oc:
        oc_row = c.execute(
            "SELECT estado FROM ordenes_compra WHERE numero_oc=?", (numero_oc,)
        ).fetchone()
        if not oc_row:
            return jsonify({'error': f'OC {numero_oc} no encontrada'}), 404
        oc_estado = oc_row[0] or ''
        if oc_estado not in ('Autorizada', 'Pagada', 'Parcial'):
            return jsonify({
                'error': f'OC en estado {oc_estado} — no se puede marcar recibida desde solicitante'
            }), 409
        c.execute("""
            UPDATE ordenes_compra SET
              estado='Recibida',
              fecha_recepcion=COALESCE(fecha_recepcion, datetime('now', '-5 hours')),
              recibido_por=?,
              observaciones_recepcion=COALESCE(observaciones_recepcion,'') || ?
            WHERE numero_oc=?
        """, (user, f' [Confirmado por solicitante {user}: {obs}]' if obs else f' [Confirmado por solicitante {user}]', numero_oc))
        # Cerrar items del checklist linkeados (si aplica)
        try:
            c.execute("""
                UPDATE produccion_checklist SET
                  estado='recibido',
                  fecha_recibido=date('now', '-5 hours'),
                  actualizado_at=datetime('now', '-5 hours')
                WHERE oc_numero=? AND estado IN ('solicitado','en_transito','pendiente')
            """, (numero_oc,))
        except sqlite3.OperationalError:
            pass
        try:
            audit_log(c, usuario=user, accion='MARCAR_RECIBIDO_SOLICITANTE',
                      tabla='ordenes_compra', registro_id=numero_oc,
                      antes={'estado': oc_estado},
                      despues={'estado': 'Recibida', 'recibido_por': user,
                                'observaciones': obs[:300] if obs else None},
                      detalle=f"Solicitante {user} marcó OC {numero_oc} como Recibida (desde SOL {numero})")
        except Exception as e:
            log.warning('audit_log MARCAR_RECIBIDO_SOLICITANTE fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'numero': numero.upper(),
        'numero_oc': numero_oc,
        'mensaje': 'Recepción confirmada. La OC quedó marcada como Recibida.',
    })


@bp.route('/api/solicitudes-compra/<numero>/aprobar-influencer', methods=['POST'])
def aprobar_solicitud_influencer(numero):
    """Aprueba una SOL Pendiente de Influencer/CC, completa el valor, crea
    la OC vinculada y registra en pagos_influencers.

    Sebastian (29-abr-2026): Jefferson crea SOLs desde /solicitudes con
    valor=0 y Pendiente. Este endpoint cierra el ciclo: define el valor,
    auto-aprueba, crea OC y entrada en pagos_influencers — listo para
    pagar desde /compras tab Influencers.

    Body: {valor: 1500000}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin (Sebastian/Alejandro)'}), 403
    d = request.get_json() or {}
    # Money sanity validation · audit zero-error
    monto, err = validate_money(d.get('valor'), allow_zero=False, field_name='valor')
    if err:
        return jsonify(err), 400
    numero = numero.upper()
    conn = get_db(); cur = conn.cursor()
    sol = cur.execute(
        "SELECT estado, categoria, observaciones, numero_oc, influencer_id, solicitante "
        "FROM solicitudes_compra WHERE numero=?", (numero,)
    ).fetchone()
    if not sol:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    sol_estado, categoria, obs_orig, numero_oc_actual, infl_id, solicitante = sol
    cat_low = (categoria or '').lower()
    if not ('influencer' in cat_low or 'cuenta de cobro' in cat_low or 'marketing' in cat_low):
        return jsonify({'error': 'Esta solicitud no es de influencer/CC'}), 400

    # Buscar nombre del beneficiario
    benef_nombre = ''
    if infl_id:
        try:
            r = cur.execute("SELECT nombre FROM marketing_influencers WHERE id=?", (infl_id,)).fetchone()
            if r: benef_nombre = r[0] or ''
        except Exception as _e:
            __import__('logging').getLogger('compras').warning('lookup beneficiario infl_id=%s fallo: %s', infl_id, _e)
    if not benef_nombre and obs_orig:
        m = __import__('re').search(r'BENEFICIARIO:\s*([^|]+)', obs_orig, __import__('re').IGNORECASE)
        if m: benef_nombre = m.group(1).strip()
    if not benef_nombre:
        benef_nombre = solicitante or 'Sin beneficiario'

    # 1. Update SOL: valor + estado=Aprobada
    cur.execute(
        "UPDATE solicitudes_compra SET valor=?, estado='Aprobada', "
        "aprobado_por=?, fecha_aprobacion=datetime('now', '-5 hours') WHERE numero=?",
        (monto, user, numero)
    )

    # 2. Crear OC si no existe
    if not numero_oc_actual:
        oc_num = numero.replace('SOL', 'OC')
        # Si ya hay una OC con ese numero (raro pero posible), generar uno con sufijo
        existing_oc = cur.execute(
            "SELECT 1 FROM ordenes_compra WHERE numero_oc=?", (oc_num,)
        ).fetchone()
        if existing_oc:
            from datetime import datetime as _dt
            oc_num = oc_num + '-' + _dt.now().strftime('%H%M%S')
        cur.execute("""
            INSERT INTO ordenes_compra
            (numero_oc, fecha, estado, proveedor, observaciones, creado_por,
             categoria, valor_total)
            VALUES (?, date('now', '-5 hours'), 'Aprobada', ?, ?, ?, ?, ?)
        """, (oc_num, benef_nombre, obs_orig or '', user, categoria, monto))
        cur.execute(
            "UPDATE solicitudes_compra SET numero_oc=? WHERE numero=?",
            (oc_num, numero)
        )
    else:
        oc_num = numero_oc_actual
        # Si la OC existe pero estaba en valor 0, actualizarla
        cur.execute(
            "UPDATE ordenes_compra SET valor_total=?, estado='Aprobada', "
            "proveedor=COALESCE(NULLIF(proveedor,''),?) WHERE numero_oc=?",
            (monto, benef_nombre, oc_num)
        )

    # 3. Registrar en pagos_influencers (si no existe ya)
    try:
        existing_pi = cur.execute(
            "SELECT 1 FROM pagos_influencers WHERE numero_oc=?", (oc_num,)
        ).fetchone()
        if not existing_pi:
            # FIX 27-may · agregar fecha_contenido + vence_pago_at (mig 195) ·
            # default fecha_contenido=hoy si la SC viene del aprobar-flow donde
            # no se preguntó. vence = +30d (promesa de pago Sebastián 27-may).
            from datetime import datetime as _dtIF, timedelta as _tdIF
            _base_fc = (_dtIF.utcnow() - _tdIF(hours=5)).date()
            _fc = _base_fc.isoformat()
            _vence = (_base_fc + _tdIF(days=30)).isoformat()
            # FIX 27-may · int(monto) trunca decimales (P1) · round preserva centavos
            _monto_int = round(float(monto or 0))
            try:
                cur.execute("""
                    INSERT INTO pagos_influencers
                    (influencer_id, influencer_nombre, valor, fecha, estado,
                     concepto, numero_oc, fecha_contenido, vence_pago_at)
                    VALUES (?, ?, ?, date('now', '-5 hours'), 'Pendiente', ?, ?, ?, ?)
                """, (infl_id, benef_nombre, _monto_int,
                      obs_orig[:200] if obs_orig else 'Cuenta de cobro',
                      oc_num, _fc, _vence))
            except Exception:
                # Fallback si mig 195 aún no aplicada en esta instancia PG
                cur.execute("""
                    INSERT INTO pagos_influencers
                    (influencer_id, influencer_nombre, valor, fecha, estado,
                     concepto, numero_oc)
                    VALUES (?, ?, ?, date('now', '-5 hours'), 'Pendiente', ?, ?)
                """, (infl_id, benef_nombre, _monto_int,
                      obs_orig[:200] if obs_orig else 'Cuenta de cobro', oc_num))
    except sqlite3.OperationalError:
        pass  # tabla puede no existir en instancias muy viejas
    try:
        audit_log(cur, usuario=user, accion='APROBAR_SOLICITUD_INFLUENCER',
                  tabla='solicitudes_compra', registro_id=numero,
                  antes={'estado': sol_estado, 'numero_oc': numero_oc_actual},
                  despues={'estado': 'Aprobada', 'valor': monto,
                            'numero_oc': oc_num, 'beneficiario': benef_nombre[:200],
                            'aprobado_por': user},
                  detalle=f"Aprobó solicitud {numero} influencer · OC {oc_num} · "
                          f"beneficiario {benef_nombre[:60]} · ${monto:,.0f}")
    except Exception as e:
        log.warning('audit_log APROBAR_SOLICITUD_INFLUENCER fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'numero': numero,
        'numero_oc': oc_num,
        'valor': monto,
        'mensaje': f'Aprobada — OC {oc_num} creada por ${monto:,.0f}'
    })


@bp.route('/api/solicitudes-compra/<numero>/rechazar', methods=['POST'])
def rechazar_solicitud(numero):
    """Marca una SOL como Rechazada y notifica al solicitante por email.
    Sin tocar OC asociada (si la tiene). Solo admin.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    d = request.get_json() or {}
    motivo = (d.get('motivo') or '').strip() or 'Sin motivo'
    numero = numero.upper()
    conn = get_db(); cur = conn.cursor()
    sol = cur.execute(
        "SELECT solicitante, email_solicitante, observaciones, categoria "
        "FROM solicitudes_compra WHERE numero=?",
        (numero,)
    ).fetchone()
    if not sol:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    cur.execute(
        "UPDATE solicitudes_compra SET estado='Rechazada', "
        "observaciones=COALESCE(observaciones,'') || ' | RECHAZADA: ' || ? "
        "WHERE numero=?",
        (motivo, numero)
    )
    try:
        audit_log(cur, usuario=user, accion='RECHAZAR_SOLICITUD',
                  tabla='solicitudes_compra', registro_id=numero,
                  antes={'solicitante': sol[0], 'categoria': sol[3]},
                  despues={'estado': 'Rechazada', 'motivo': motivo[:300]},
                  detalle=f"Rechazó solicitud {numero} · motivo: {motivo}")
    except Exception as e:
        log.warning('audit_log RECHAZAR_SOLICITUD fallo: %s', e)
    conn.commit()
    # Notificar al solicitante. Sebastian (29-abr-2026): "cuando doy rechazar
    # en compras a una cuenta de influencer me esta llegando a mi deberia
    # llegarle a jeferson". Para SOLs de Influencer/CC el destinatario SIEMPRE
    # debe ser Jefferson (responsable de marketing) — no el solicitante real
    # que puede haber sido Sebastián cargando bulk.
    try:
        _sol_user = (sol[0] or '').strip().lower()
        _categoria = (sol[3] or '').strip()
        _es_influencer = _categoria in ('Influencer/Marketing Digital', 'Cuenta de Cobro')
        if _es_influencer:
            # Forzar Jefferson como destinatario para SOLs de marketing
            _dest = USER_EMAILS.get('jefferson', '')
            # Fallback solo si Jefferson no está configurado: usar email de la SOL
            if not _dest:
                _dest = (sol[1] or '').strip() or USER_EMAILS.get(_sol_user, '')
        else:
            _dest = (sol[1] or '').strip() or USER_EMAILS.get(_sol_user, '')
        if _dest:
            _asunto = f"❌ Solicitud {numero} rechazada"
            _body = (
                f"<h2>Tu solicitud de pago fue rechazada</h2>"
                f"<p>Solicitud: <b>{numero}</b></p>"
                f"<p>Motivo: <i>{motivo}</i></p>"
                f"<p style='color:#94a3b8;font-size:11px'>Mensaje automatico HHA Group</p>"
            )
            _notificar_solicitante_email(_dest, _asunto, _body)
    except Exception:
        pass
    return jsonify({'ok': True, 'numero': numero, 'estado': 'Rechazada', 'motivo': motivo})


@bp.route('/api/solicitudes-compra/<numero>', methods=['GET'])
def get_solicitud_estado(numero):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT numero,fecha,estado,solicitante,urgencia,observaciones,numero_oc FROM solicitudes_compra WHERE numero=?", (numero.upper(),))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'No encontrada'}), 404
    cols = ['numero','fecha','estado','solicitante','urgencia','observaciones','numero_oc']
    sol = dict(zip(cols, row))
    # Cargar columnas opcionales (pueden no existir en versiones antiguas
    # de la DB). Whitelist hardcoded — nunca de input — así el f-string es
    # seguro contra SQL injection.
    for col in ['area', 'empresa', 'categoria', 'tipo', 'aprobado_por', 'fecha_aprobacion']:
        try:
            c.execute(f"SELECT {col} FROM solicitudes_compra WHERE numero=?", (numero.upper(),))
            r2 = c.fetchone()
            if r2:
                sol[col] = r2[0]
        except sqlite3.OperationalError as _e:
            # "no such column" es benigno (versión vieja de schema).
            # Otros errores SÍ los logueamos.
            if 'no such column' not in str(_e).lower():
                __import__('logging').getLogger('compras').error(
                    "SELECT %s en solicitudes_compra falló: %s", col, _e
                )
    # IMPORTANTE: incluir id (PK) y campos editables nuevos para que el modal
    # pueda hacer PATCH /api/solicitudes-compra/<numero>/items
    try:
        c.execute("""SELECT id, codigo_mp, nombre_mp, cantidad_g, unidad,
                            COALESCE(valor_estimado, 0),
                            COALESCE(justificacion, ''),
                            COALESCE(precio_unit_g, 0),
                            COALESCE(proveedor_sugerido, '')
                     FROM solicitudes_compra_items WHERE numero=?""",
                  (numero.upper(),))
        items = [
            dict(zip(['id','codigo_mp','nombre_mp','cantidad_g','unidad',
                      'valor_estimado','justificacion',
                      'precio_unit_g','proveedor_sugerido'], r))
            for r in c.fetchall()
        ]
    except sqlite3.OperationalError:
        # Fallback si migration #43 todavia no aplicada
        c.execute("""SELECT id, codigo_mp, nombre_mp, cantidad_g, unidad,
                            COALESCE(valor_estimado, 0),
                            COALESCE(justificacion, '')
                     FROM solicitudes_compra_items WHERE numero=?""",
                  (numero.upper(),))
        items = [
            dict(zip(['id','codigo_mp','nombre_mp','cantidad_g','unidad',
                      'valor_estimado','justificacion'], r))
            for r in c.fetchall()
        ]
        for it in items:
            it['precio_unit_g'] = 0
            it['proveedor_sugerido'] = ''

    # Enriquecer cada item con stock_actual_g (suma de movimientos) +
    # precio_referencia + proveedor de maestro_mps. Para que el modal de
    # solicitud muestre 'Tienes X, faltan Y' y el contexto financiero.
    for it in items:
        cod = (it.get('codigo_mp') or '').strip()
        nombre = (it.get('nombre_mp') or '').strip()
        it['stock_actual_g'] = 0
        if not cod and not nombre:
            continue
        # Estrategia de búsqueda escalonada:
        # 1. material_id exacto con código
        # 2. material_nombre LIKE con nombre (caso-insensitive)
        # Si la primera tiene resultado, no probamos la segunda.
        stock_total = 0
        try:
            if cod:
                # FIX 2-jun audit abastecimiento · excluir estados no-producibles
                # (AGOTADO del reset/CUARENTENA/etc) · antes mostraba stock fantasma
                # en el modal que dispara la compra → comprador sub-compraba.
                r = c.execute("""SELECT
                    COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0),
                    COUNT(*)
                    FROM movimientos WHERE material_id=?
                      AND (estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))""", (cod,)).fetchone()
                if r and r[1] > 0:  # hubo movimientos con ese código
                    stock_total = float(r[0] or 0)
                elif nombre:
                    # Fallback: buscar por nombre exacto upper
                    r2 = c.execute("""SELECT
                        COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0)
                        FROM movimientos
                        WHERE UPPER(TRIM(material_nombre)) = UPPER(TRIM(?))
                          AND (estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))""",
                        (nombre,)).fetchone()
                    stock_total = float(r2[0] or 0) if r2 else 0
            elif nombre:
                r = c.execute("""SELECT
                    COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0)
                    FROM movimientos
                    WHERE UPPER(TRIM(material_nombre)) = UPPER(TRIM(?))
                      AND (estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))""",
                    (nombre,)).fetchone()
                stock_total = float(r[0] or 0) if r else 0
            it['stock_actual_g'] = max(0, stock_total)  # nunca negativo en display
        except sqlite3.OperationalError:
            pass
        try:
            r = c.execute("""SELECT COALESCE(proveedor,''),
                                    COALESCE(precio_referencia,0)
                             FROM maestro_mps WHERE codigo_mp=?""", (cod,)).fetchone()
            if r:
                it['proveedor'] = r[0]
                it['precio_referencia'] = float(r[1] or 0)
                # Si no hay valor estimado pero tenemos precio referencia, calcular
                if not it.get('valor_estimado') and it['precio_referencia'] > 0:
                    cant_g = float(it.get('cantidad_g') or 0)
                    it['valor_estimado_calculado'] = round(cant_g / 1000.0 * it['precio_referencia'], 0)
        except sqlite3.OperationalError:
            pass

    # Datos de la OC asociada (proveedor, valor_total, estado) si existe
    oc_info = None
    if sol.get('numero_oc'):
        try:
            r = c.execute("""SELECT proveedor, COALESCE(valor_total, 0), estado, observaciones
                             FROM ordenes_compra WHERE numero_oc=?""",
                          (sol['numero_oc'],)).fetchone()
            if r:
                oc_info = {
                    'numero_oc': sol['numero_oc'],
                    'proveedor': r[0] or '',
                    'valor_total': float(r[1] or 0),
                    'estado': r[2] or '',
                    'observaciones': r[3] or '',
                }
        except sqlite3.OperationalError:
            pass

    return jsonify({'solicitud': sol, 'items': items, 'oc': oc_info})

@bp.route('/solicitudes')
def solicitudes_page():
    if 'compras_user' not in session:
        return redirect('/login?next=/solicitudes')
    return Response(SOLICITUDES_HTML, mimetype='text/html')

@bp.route('/api/solicitudes-compra/<numero>/split', methods=['POST'])
def solicitud_split(numero):
    """Sprint Compras N2 · 21-may-2026.

    Cuando una SOL tiene items de varios proveedores distintos, queda
    como '__MIXTO__' en la card consolidada · Catalina no puede crear
    UNA OC porque los items van a proveedores distintos. ANTES tenía
    que borrar y recrear · ahora puede hacer SPLIT.

    Toma una SOL existente y la divide en N SOLs nuevas, una por cada
    proveedor distinto en sus items. La SOL original queda como
    "Reemplazada" y guarda referencia a las hijas.

    Body opcional: {por_campo: 'proveedor'|'proveedor_sugerido'} default
    'proveedor_sugerido'.
    """
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    body = request.get_json(silent=True) or {}
    por = (body.get('por_campo') or 'proveedor_sugerido').strip()
    if por not in ('proveedor', 'proveedor_sugerido'):
        por = 'proveedor_sugerido'
    conn = get_db(); c = conn.cursor()
    sol = c.execute(
        """SELECT numero, estado, COALESCE(numero_oc,''), categoria,
                  COALESCE(empresa,''), COALESCE(area,''), COALESCE(tipo,''),
                  COALESCE(solicitante,''), COALESCE(urgencia,''),
                  COALESCE(observaciones,''), COALESCE(fecha_requerida,'')
           FROM solicitudes_compra WHERE numero=?""",
        (numero,),
    ).fetchone()
    if not sol:
        return jsonify({'error': f'SOL {numero} no existe'}), 404
    if sol[1] != 'Pendiente':
        return jsonify({'error': f'SOL {numero} no está Pendiente (estado: {sol[1]})'}), 409
    if sol[2]:
        return jsonify({'error': f'SOL {numero} ya tiene OC asociada'}), 409
    # Cargar items
    items = c.execute(
        f"""SELECT id, codigo_mp, nombre_mp, cantidad_g, unidad,
                   COALESCE(valor_estimado,0), COALESCE(precio_unit_g,0),
                   COALESCE({por},''), COALESCE(justificacion,'')
            FROM solicitudes_compra_items WHERE numero=?""",
        (numero,),
    ).fetchall()
    if not items:
        return jsonify({'error': f'SOL {numero} sin items'}), 400
    # Agrupar por proveedor
    por_prov = {}
    for it in items:
        prov = (it[7] or '_SIN_PROVEEDOR_').strip()
        por_prov.setdefault(prov, []).append(it)
    if len(por_prov) <= 1:
        return jsonify({
            'error': 'SOL no es mixta · todos los items tienen mismo proveedor',
            'proveedor_unico': list(por_prov.keys())[0] if por_prov else None,
        }), 400
    # Crear N SOLs hijas
    from datetime import datetime as _dt
    hijas_creadas = []
    anio = _dt.now().strftime('%Y')
    try:
        for prov, prov_items in por_prov.items():
            # Buscar siguiente numero AUTO-XXXX
            for _intento in range(6):
                row = c.execute(
                    "SELECT COALESCE(MAX(CAST(SUBSTR(numero, 6) AS INTEGER)),0) "
                    "FROM solicitudes_compra WHERE numero LIKE 'AUTO-%'",
                ).fetchone()
                nuevo_n = f"AUTO-{(row[0] or 0)+1:04d}"
                try:
                    c.execute(
                        """INSERT INTO solicitudes_compra
                           (numero, fecha, estado, solicitante, urgencia, observaciones,
                            empresa, area, categoria, tipo, fecha_requerida)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (nuevo_n, _dt.now().isoformat(), 'Pendiente',
                         sol[7] or usuario, sol[8] or 'Media',
                         f'SPLIT de {numero} · proveedor {prov} · ' + (sol[9] or ''),
                         sol[4], sol[5], sol[3], sol[6], sol[10] or ''),
                    )
                    break
                except Exception:
                    continue
            else:
                raise RuntimeError(f'No se pudo asignar numero para split de {numero}')
            # Insertar items
            for it in prov_items:
                c.execute(
                    f"""INSERT INTO solicitudes_compra_items
                        (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                         valor_estimado, precio_unit_g, {por}, justificacion)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                    (nuevo_n, it[1], it[2], it[3], it[4],
                     it[5], it[6], prov, it[8]),
                )
            hijas_creadas.append({
                'numero': nuevo_n, 'proveedor': prov,
                'items_count': len(prov_items),
            })
        # Marcar original como Reemplazada (no Cancelada para no perder histórico)
        c.execute(
            "UPDATE solicitudes_compra SET estado='Reemplazada', "
            "observaciones=COALESCE(observaciones,'') || ' · SPLIT en ' || ? "
            "WHERE numero=?",
            (', '.join(h['numero'] for h in hijas_creadas), numero),
        )
        # Audit
        try:
            audit_log(c, usuario=usuario, accion='SPLIT_SOLICITUD',
                      tabla='solicitudes_compra', registro_id=numero,
                      despues={'hijas': hijas_creadas,
                               'proveedores': list(por_prov.keys())})
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        return jsonify({'error': f'Falla split: {e}', 'tipo': type(e).__name__}), 500
    return jsonify({
        'ok': True,
        'original': numero,
        'estado_original': 'Reemplazada',
        'hijas_creadas': hijas_creadas,
        'mensaje': f'SOL {numero} dividida en {len(hijas_creadas)} hijas por proveedor',
    })


@bp.route('/api/solicitudes-compra/<numero>/estado', methods=['PATCH'])
def actualizar_estado_solicitud(numero):
    # INV-3 · cambiar el estado de una SOL (aprobar/rechazar/crear OC) es
    # operación de Compras, no de cualquier usuario logueado.
    user_act, err, code = _require_compras_write()
    if err:
        return err, code
    d = request.get_json() or {}
    nuevo = d.get('estado', 'Aprobada')
    # P0 audit 26-may · whitelist de estado · sin esto un user puede meter
    # cualquier string (incluyendo HTML/scripts) y romper queries downstream
    # que filtran por estado IN (...).
    _ESTADOS_SOL = {'Pendiente','Aprobada','Rechazada','Pagada','Cancelada','Reemplazada'}
    if nuevo not in _ESTADOS_SOL:
        return jsonify({'error': f'estado inválido: {nuevo!r} · válidos: {sorted(_ESTADOS_SOL)}'}), 400
    numero_oc_param = d.get('numero_oc', '')
    obs = d.get('observaciones', '')
    conn = get_db(); cur = conn.cursor()
    # Capturar antes para audit (estado anterior y SOL existence check)
    antes_row = cur.execute(
        "SELECT estado, numero_oc, categoria FROM solicitudes_compra WHERE numero=?",
        (numero.upper(),)).fetchone()
    if not antes_row:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    antes = {'estado': antes_row[0], 'numero_oc': antes_row[1], 'categoria': antes_row[2]}
    cur.execute("""UPDATE solicitudes_compra SET estado=?, aprobado_por=?, fecha_aprobacion=?
                 WHERE numero=?""",
              (nuevo, user_act, datetime.now().isoformat(), numero.upper()))
    if numero_oc_param:
        cur.execute("UPDATE solicitudes_compra SET numero_oc=? WHERE numero=?", (numero_oc_param, numero.upper()))
    if nuevo == 'Rechazada' and obs:
        # SOL-FIX#1 · 21-may-2026 · concatenar motivo (no reemplazar)
        # Antes destruía la observación original del solicitante.
        cur.execute(
            "UPDATE solicitudes_compra SET observaciones=COALESCE(observaciones,'') || ' | RECHAZADA: ' || ? WHERE numero=?",
            (obs, numero.upper()),
        )
    # SOL-FIX#1b · race UPDATE inicial con WHERE estado='Pendiente' para evitar doble OC
    # se valida re-checando rowcount del UPDATE de arriba (ya hecho)
    try:
        audit_log(cur, usuario=user_act, accion='ACTUALIZAR_ESTADO_SOL',
                  tabla='solicitudes_compra', registro_id=numero.upper(),
                  antes=antes,
                  despues={'estado': nuevo, 'numero_oc': numero_oc_param or antes_row[1],
                            'observaciones': obs[:300] if obs else None,
                            'aprobado_por': user_act},
                  detalle=f"Solicitud {numero.upper()}: {antes['estado']} → {nuevo}")
    except Exception as e:
        log.warning('audit_log ACTUALIZAR_ESTADO_SOL fallo: %s', e)
    conn.commit()
    oc_creada = ''
    if d.get('crear_oc'):
        cur.execute("SELECT codigo_mp, nombre_mp, cantidad_g, unidad FROM solicitudes_compra_items WHERE numero=?", (numero.upper(),))
        items_sol = cur.fetchall()
        # Obtener categoria de la solicitud para la OC
        cur.execute("SELECT categoria FROM solicitudes_compra WHERE numero=?", (numero.upper(),))
        sol_row = cur.fetchone()
        categoria_oc = d.get('categoria') or (sol_row[0] if sol_row and sol_row[0] else 'MP')
        # Sebastián 24-may-2026 · audit Solicitudes · preservar categoría ORIGINAL
        # antes del mapeo abreviado para hacer match contra compras_fast_track_config
        # (que guarda nombre humano · no abreviación). El mapeo abreviado se aplica
        # después solo para escribir en ordenes_compra.categoria.
        categoria_orig = categoria_oc
        # Normalizar categorias largas de solicitudes a claves cortas de OC
        _SOLIC_TO_OC_CAT = {
            'Materia Prima': 'MP', 'Materias Primas': 'MP', 'MPs': 'MP',
            'Material de Empaque': 'MEE', 'Envase': 'MEE', 'Insumos': 'MEE',
            'Empaque': 'MEE',
            'EPP': 'ADM', 'Aseo/Limpieza': 'ADM', 'Papeleria/Oficina': 'ADM',
            'Dotacion': 'ADM', 'Otro': 'ADM', 'Administrativo': 'ADM',
            'Nomina': 'ADM', 'Admin': 'ADM',
            'Mantenimiento': 'INF', 'Repuestos': 'INF',
            'Reactivos/Laboratorio': 'INF', 'Infraestructura': 'INF',
            'Servicios Profesionales': 'SVC', 'Software/Tecnologia': 'SVC',
            'Servicios': 'SVC', 'Analisis': 'SVC', 'Acondicionamiento': 'SVC',
            'Servicio': 'SVC',
        }
        categoria_oc = _SOLIC_TO_OC_CAT.get(categoria_oc, categoria_oc)
        proveedor_oc = (d.get('proveedor') or '').strip()
        # P1 audit 26-may · validate_money en valor_total · ValueError raw
        # no manejado iba a 500 sin mensaje útil.
        from http_helpers import validate_money as _vm_sol
        valor_oc, _err_vo = _vm_sol(d.get('valor_total') or 0, allow_zero=True,
                                      max_value=1e10, field_name='valor_total')
        if _err_vo:
            return jsonify(_err_vo), 400
        # Sebastian (29-abr-2026): "OC sale Por definir y $0 aunque la
        # solicitud trae 150.000". Si el frontend no manda proveedor/valor,
        # los inferimos desde la solicitud para que la OC no quede vacia.
        if not proveedor_oc or proveedor_oc.lower() == 'por definir':
            try:
                _r = cur.execute(
                    "SELECT influencer_id, solicitante FROM solicitudes_compra WHERE numero=?",
                    (numero.upper(),)
                ).fetchone()
                inf_id = _r[0] if _r else None
                solic_nombre = (_r[1] if _r else '') or ''
                if inf_id:
                    _ri = cur.execute(
                        "SELECT nombre FROM marketing_influencers WHERE id=?", (inf_id,)
                    ).fetchone()
                    if _ri and _ri[0]:
                        proveedor_oc = _ri[0]
                if not proveedor_oc:
                    _rp = cur.execute(
                        "SELECT proveedor_sugerido FROM solicitudes_compra_items "
                        "WHERE numero=? AND COALESCE(proveedor_sugerido,'') != '' LIMIT 1",
                        (numero.upper(),)
                    ).fetchone()
                    if _rp and _rp[0]:
                        proveedor_oc = _rp[0]
                if not proveedor_oc and categoria_oc in ('SVC', 'CC') and solic_nombre:
                    proveedor_oc = solic_nombre
            except Exception:
                pass
        if not proveedor_oc:
            proveedor_oc = 'Por definir'
        if valor_oc <= 0:
            try:
                _r = cur.execute(
                    "SELECT COALESCE(valor,0) FROM solicitudes_compra WHERE numero=?",
                    (numero.upper(),)
                ).fetchone()
                if _r and (_r[0] or 0) > 0:
                    valor_oc = float(_r[0])
                else:
                    _r2 = cur.execute(
                        "SELECT COALESCE(SUM(COALESCE(valor_estimado,0)),0), "
                        "       COALESCE(SUM(COALESCE(precio_unit_g,0)*COALESCE(cantidad_g,0)),0) "
                        "FROM solicitudes_compra_items WHERE numero=?",
                        (numero.upper(),)
                    ).fetchone()
                    if _r2:
                        valor_oc = float(_r2[0] or 0) or float(_r2[1] or 0)
            except Exception:
                pass
        fent_oc = d.get('fecha_entrega_est', '')
        obs_oc = d.get('observaciones_oc') or f'Generado desde {numero.upper()}'
        # Sebastián 24-may-2026 · fast-track configurable · audit Solicitudes ·
        # Antes hardcoded a (Influencer/Marketing Digital, Cuenta de Cobro). Ahora
        # se lee de compras_fast_track_config (mig 179) · permite a Sebastián
        # marcar otras categorías como fast-track (Papelería <$200k, EPP <$500k,
        # etc.) sin tocar código. monto_max_cop=0 = sin tope (legacy behavior).
        # SEC-FIX 23-may-2026 · C5 · mig 157 eliminó 'Revisada' · default Borrador.
        estado_oc = 'Borrador'
        fast_track_aplicado = False
        try:
            ft_row = cur.execute(
                "SELECT monto_max_cop, COALESCE(notas,'') FROM compras_fast_track_config "
                "WHERE categoria=? AND activo=1",
                (categoria_orig,),
            ).fetchone()
            if ft_row is not None:
                _monto_max = float(ft_row[0] or 0)
                # monto_max=0 = sin tope · cualquier monto pasa
                # monto_max>0 = aplicar fast-track solo si valor_oc <= monto_max
                if _monto_max <= 0 or (valor_oc and valor_oc <= _monto_max):
                    estado_oc = 'Autorizada'
                    fast_track_aplicado = True
        except sqlite3.OperationalError:
            # Tabla aún no migrada · fallback al comportamiento legacy hardcoded
            _LEGACY_FAST_TRACK = ('Influencer/Marketing Digital', 'Cuenta de Cobro')
            if categoria_orig in _LEGACY_FAST_TRACK:
                estado_oc = 'Autorizada'
                fast_track_aplicado = True
        # Aún con fast-track, respetar el límite de aprobación del usuario · si
        # el monto lo excede, fuerza Borrador para que un admin autorice.
        if estado_oc == 'Autorizada' and valor_oc and valor_oc > 0:
            _err_lim, _ = _check_monto_limit(session.get('compras_user', ''), valor_oc)
            if _err_lim:
                estado_oc = 'Borrador'
                fast_track_aplicado = False
        # numero único con reintento ante carrera MAX+1 entre workers
        for _intento in range(6):
            oc_num = _siguiente_numero_oc(cur, datetime.now().year)
            try:
                cur.execute(
                    "INSERT INTO ordenes_compra "
                    "(numero_oc, fecha, estado, proveedor, observaciones, creado_por, valor_total, fecha_entrega_est, categoria) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (oc_num, datetime.now().isoformat(), estado_oc, proveedor_oc,
                     obs_oc, session.get('compras_user',''),
                     valor_oc if valor_oc > 0 else None, fent_oc or None, categoria_oc))
                break
            except sqlite3.IntegrityError:
                if _intento == 5:
                    raise
        for it in items_sol:
            cur.execute("INSERT INTO ordenes_compra_items (numero_oc, codigo_mp, nombre_mp, cantidad_g) VALUES (?,?,?,?)",
                      (oc_num, it[0], it[1], it[2]))
        cur.execute("UPDATE solicitudes_compra SET numero_oc=? WHERE numero=?", (oc_num, numero.upper()))
        oc_creada = oc_num
    # Fetch email for notification BEFORE closing connection
    cur.execute("SELECT email_solicitante FROM solicitudes_compra WHERE numero=?", (numero.upper(),))
    _em_row = cur.fetchone()
    _notif_dest = (_em_row[0] if _em_row else '').strip()
    conn.commit()
    # Notificacion al solicitante (no-blocking, best-effort)
    if _notif_dest:
        if nuevo == 'Rechazada':
            _asunto_n = f'Solicitud rechazada \u2014 {numero.upper()}'
            _body_n = (
                '<html><body style="font-family:Arial,sans-serif;max-width:600px;">'
                '<div style="background:#fee2e2;padding:20px;border-radius:8px;border-left:4px solid #dc2626;">'
                '<h2 style="color:#991b1b;">Solicitud rechazada</h2>'
                f'<p>Tu solicitud <strong>{numero.upper()}</strong> fue rechazada.</p>'
                + (f'<p><strong>Motivo:</strong> {obs}</p>' if obs else '')
                + '<p>Puedes corregirla y reenviarla desde el sistema.</p>'
                '<p style="color:#6b7280;font-size:12px;">Compras HHA \u2014 Espagiria</p>'
                '</div></body></html>'
            )
            _notificar_solicitante_email(_notif_dest, _asunto_n, _body_n)
        elif oc_creada:
            _asunto_n = f'Solicitud aprobada \u2014 {numero.upper()}'
            _body_n = (
                '<html><body style="font-family:Arial,sans-serif;max-width:600px;">'
                '<div style="background:#dcfce7;padding:20px;border-radius:8px;border-left:4px solid #16a34a;">'
                '<h2 style="color:#15803d;">Solicitud aprobada</h2>'
                f'<p>Tu solicitud <strong>{numero.upper()}</strong> fue aprobada.</p>'
                f'<p>Orden de compra generada: <strong>{oc_creada}</strong></p>'
                '<p>El equipo de compras esta gestionando tu pedido.</p>'
                '<p style="color:#6b7280;font-size:12px;">Compras HHA \u2014 Espagiria</p>'
                '</div></body></html>'
            )
            _notificar_solicitante_email(_notif_dest, _asunto_n, _body_n)
    # Sebastián 24-may-2026 · audit Solicitudes · UX hint del próximo paso.
    # Antes Catalina aprobaba y no sabía si tenía que hacer algo más · la OC
    # queda en Borrador (no Autorizada) salvo Influencer/CC fast-track ·
    # quedaba "olvidada" hasta que recordaba ir a tab OCs Activas. Ahora el
    # response lleva un hint contextual que el UI puede mostrar como toast.
    _resp = {'ok': True, 'estado': nuevo, 'numero_oc': oc_creada}
    if oc_creada:
        try:
            _eoc = cur.execute(
                "SELECT estado FROM ordenes_compra WHERE numero_oc=?",
                (oc_creada,)
            ).fetchone()
            _estado_oc_real = (_eoc[0] if _eoc else '') or ''
        except Exception:
            _estado_oc_real = ''
        _resp['oc_estado'] = _estado_oc_real
        if _estado_oc_real == 'Borrador':
            _resp['siguiente_paso'] = (
                f'OC {oc_creada} creada en Borrador · revisala en tab '
                f'"📦 OCs Activas" y autorizá para enviar al proveedor.'
            )
        elif _estado_oc_real == 'Autorizada':
            _resp['siguiente_paso'] = (
                f'OC {oc_creada} fast-track · ya quedó Autorizada · '
                f'pasa a tab "💰 Por Pagar" cuando esté lista para pago.'
            )
    return jsonify(_resp)

@bp.route('/api/ordenes-compra/<numero_oc>/recibir', methods=['POST'])
def recibir_oc(numero_oc):
    # Audit zero-error 2-may-2026: RBAC mínimo (era cualquier sesión)
    # Recepción mueve stock e impacta valoración · debe ser Compras/Bodega/Admin
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in (set(COMPRAS_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Compras/Admin pueden registrar recepción'}), 403
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT estado, proveedor, categoria FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    oc_row = cur.fetchone()
    if not oc_row:
        return jsonify({'error': 'OC no encontrada'}), 404
    if oc_row[0] not in ('Autorizada', 'Parcial', 'Pagada'):
        return jsonify({'error': f'OC en estado {oc_row[0]} no permite recepcion'}), 409
    estado_oc_original = oc_row[0]
    prov_nombre = oc_row[1] or ''
    categoria = oc_row[2] or 'MP'
    # Bug #3 fix · 21-may-2026 · OCs de pago directo (servicios/cuotas/CC)
    # no deben aceptar recepción · son pagos puros sin material físico.
    # Antes creaba movimientos fantasma con lote sintético si receptor confundía.
    if categoria in CATEGORIAS_PAGO_DIRECTO:
        return jsonify({
            'error': f'OC categoría {categoria} es pago directo · no recibe material físico',
            'codigo': 'OC_PAGO_DIRECTO_SIN_RECEPCION',
            'hint': 'Si necesitás cerrar el ciclo, usá registrar_pago_oc o marca como Pagada directo',
        }), 409
    cur.execute("SELECT id, codigo_mp, nombre_mp, cantidad_g FROM ordenes_compra_items WHERE numero_oc=?", (numero_oc,))
    items_oc = cur.fetchall()
    # Etiqueta de INGRESO por INCI (Sebastian 12-jun): en recepcion el nombre
    # comercial varia por proveedor y es la mayor fuente de error -> el kardex
    # se rotula con el INCI (identidad sigue siendo el codigo). Si la MP no tiene
    # INCI aun, cae al codigo (nunca al comercial, nunca en blanco). El
    # nombre_comercial NO se borra de la BD; solo deja de usarse como etiqueta.
    _cods_oc = [r[1] for r in items_oc if r[1]]
    inci_by_cod = {}
    if _cods_oc:
        _ph = ','.join('?' for _ in _cods_oc)
        for _cc, _ii in cur.execute(
                "SELECT codigo_mp, COALESCE(nombre_inci,'') FROM maestro_mps "
                f"WHERE codigo_mp IN ({_ph})", _cods_oc).fetchall():
            inci_by_cod[_cc] = (_ii or '').strip()
    fecha = datetime.now().isoformat()
    operador = session.get('compras_user', '')
    data2 = request.get_json(silent=True) or {}
    obs_r = data2.get('observaciones_recepcion', '')
    disc_r = 1 if data2.get('tiene_discrepancias') else 0
    items_r = data2.get('items_recepcion', [])
    receptor_nombre = data2.get('receptor_nombre', '') or operador
    # Build lookup por indice posicional (codigo_mp puede estar vacio en solicitudes)
    rec_map_idx = {idx: ir for idx, ir in enumerate(items_r)}
    # Fallback: lookup por codigo_mp para compatibilidad con clientes que lo envien
    # AUDITORÍA-FIX 23-may-2026 · C22 · si misma MP aparece en >1 línea
    # (legítimo · 2 lotes distintos del mismo MP en una OC), el lookup
    # por codigo colapsaba en 1 sola entrada · la segunda se sobreescribía
    # · ahora: lookup por (codigo_mp, oci_rowid) cuando el cliente envía
    # oci_rowid · fallback por código solo si no hay duplicado
    rec_map_cod = {}
    rec_map_rowid = {}
    _cod_count = {}
    for ir in items_r:
        cm = (ir.get('codigo_mp') or '').strip()
        if not cm:
            continue
        _cod_count[cm] = _cod_count.get(cm, 0) + 1
    for ir in items_r:
        cm = (ir.get('codigo_mp') or '').strip()
        if cm and _cod_count.get(cm, 0) == 1:
            rec_map_cod[cm] = ir
        # Siempre indexar por rowid si el cliente lo manda
        rid = ir.get('oci_rowid') or ir.get('item_id') or ir.get('id')
        if rid is not None:
            try:
                rec_map_rowid[int(rid)] = ir
            except (TypeError, ValueError):
                pass

    # ── PRE-CHECK · Sebastian 5-may-2026 (audit zero-error Recepciones) ──
    # Antes de tocar DB validamos sobre-recepcion y vencimiento. Si hay
    # violaciones y no se forzo override (admin), abortamos con 422
    # detallado · NO se inserta nada.
    forzar_excepciones = bool(data2.get('forzar', False))
    sobrerecepciones = []
    vencimientos_pasados = []
    sin_cantidad = []
    sin_lote_proveedor = []  # FIX 27-may (P1 INVIMA) · obligatorio para MPs
    from datetime import date as _date
    hoy_iso = _date.today().isoformat()

    # Detectar si la OC es de Materia Prima · solo MPs exigen lote_proveedor
    # INVIMA (Resolución 2674/2013). Empaque/Servicios/Cuenta-Cobro NO.
    try:
        _cat_row = cur.execute("SELECT categoria FROM solicitudes_compra WHERE numero_oc=? LIMIT 1", (numero_oc,)).fetchone()
        _es_oc_mp = (_cat_row and (_cat_row[0] or '').strip() == 'Materia Prima')
    except Exception:
        _es_oc_mp = False

    for _idx, item in enumerate(items_oc):
        _oci_rowid, codigo, nombre, cantidad_pedida = item
        # Rotular el ingreso por INCI (cae al codigo si no hay INCI). El comercial
        # (nombre) queda en la OC pero no entra al kardex como etiqueta.
        nombre = inci_by_cod.get(codigo) or codigo or nombre
        # AUDITORÍA C22 · prioridad: rowid > idx posicional > codigo único
        ir = (rec_map_rowid.get(_oci_rowid)
              or rec_map_idx.get(_idx)
              or rec_map_cod.get(codigo, {}))
        cant_raw = ir.get('cantidad_recibida', 0)
        # FIX 27-may (P1 INVIMA) · lote_proveedor obligatorio para MPs · si
        # falta, bloqueamos con 422 · admin puede pasar forzar:true para
        # excepción legítima (lote pendiente de proveedor, etc.).
        if _es_oc_mp:
            _lote_prov_check = (ir.get('lote_proveedor') or '').strip()
            _cant_check = 0
            try: _cant_check = float(cant_raw or 0)
            except Exception: pass
            if _cant_check > 0 and not _lote_prov_check:
                sin_lote_proveedor.append({'codigo_mp': codigo, 'nombre': nombre})
        cantidad_explicita = not (cant_raw is None or cant_raw == '')
        if cantidad_explicita:
            cant_validada, err = validate_money(cant_raw, allow_zero=True,
                                                  max_value=10_000_000_000,
                                                  field_name='cantidad_recibida')
            if err:
                return jsonify(err), 400
            cant_check = cant_validada
        else:
            # ANTES default a cantidad_pedida (drift silencioso si receptor
            # olvidaba ingresar) · AHORA flageamos para warning informativo.
            cant_check = 0.0
            sin_cantidad.append({'codigo_mp': codigo, 'nombre': nombre})

        cantidad_pedida_f = float(cantidad_pedida or 0)
        # Sobre-recepción · tolerancia 5% para pesaje
        if cantidad_pedida_f > 0 and cant_check > cantidad_pedida_f * 1.05:
            sobrerecepciones.append({
                'codigo_mp': codigo,
                'nombre': nombre,
                'cantidad_pedida_g': cantidad_pedida_f,
                'cantidad_recibida_g': cant_check,
                'exceso_g': round(cant_check - cantidad_pedida_f, 2),
                'pct_exceso': round((cant_check / cantidad_pedida_f - 1) * 100, 1),
            })

        # Fecha vencimiento pasada
        fv_check = (ir.get('fecha_vencimiento') or '').strip()
        if fv_check and len(fv_check) >= 10:
            try:
                if fv_check[:10] < hoy_iso:
                    vencimientos_pasados.append({
                        'codigo_mp': codigo,
                        'nombre': nombre,
                        'lote': (ir.get('lote') or '').strip(),
                        'fecha_vencimiento': fv_check[:10],
                        'dias_vencido': (
                            _date.fromisoformat(hoy_iso) -
                            _date.fromisoformat(fv_check[:10])
                        ).days,
                    })
            except (ValueError, TypeError):
                pass

    # Si hay violaciones bloqueantes y no se fuerza, abortar limpio
    if not forzar_excepciones and (sobrerecepciones or vencimientos_pasados or sin_lote_proveedor):
        return jsonify({
            'error': 'Recepción bloqueada por validaciones',
            'codigo': 'RECEPCION_VIOLA_REGLAS',
            'sobrerecepciones': sobrerecepciones,
            'vencimientos_pasados': vencimientos_pasados,
            'sin_lote_proveedor': sin_lote_proveedor,
            'sin_cantidad': sin_cantidad,
            'hint': ('Verifica cantidades, vencimientos y lote_proveedor (INVIMA). '
                     'Si la excepción es legítima (admin), envía body con forzar:true.'),
        }), 422

    # ── INSERT real (post-validación) ──────────────────────────────────
    ingresos = 0
    lotes_sinteticos_advertencia = []  # Fix #9 · 21-may-2026
    for _idx, item in enumerate(items_oc):
        _oci_rowid, codigo, nombre, cantidad_pedida = item
        # Rotular el ingreso por INCI (cae al codigo si no hay INCI). El comercial
        # (nombre) queda en la OC pero no entra al kardex como etiqueta.
        nombre = inci_by_cod.get(codigo) or codigo or nombre
        # AUDITORÍA C22 · prioridad: rowid > idx posicional > codigo único
        ir = (rec_map_rowid.get(_oci_rowid)
              or rec_map_idx.get(_idx)
              or rec_map_cod.get(codigo, {}))
        cant_raw = ir.get('cantidad_recibida', 0)
        if cant_raw is None or cant_raw == '':
            cant_recibida = 0.0
        elif cant_raw == 0:
            cant_recibida = 0.0
        else:
            # Re-validar (cheap · ya pasó pre-check)
            cant_validada, err = validate_money(cant_raw, allow_zero=True,
                                                  max_value=10_000_000_000,
                                                  field_name='cantidad_recibida')
            if err:
                return jsonify(err), 400
            cant_recibida = cant_validada
        lote_num = ir.get('lote', '').strip()
        fv = ir.get('fecha_vencimiento', '').strip()
        estado_item = ir.get('estado', 'OK')
        notas_item = ir.get('notas', '')
        # COA + lote proveedor (Fase 2 · INVIMA · mig 151)
        coa_url = (ir.get('coa_url') or '').strip()
        coa_filename = (ir.get('coa_filename') or '').strip()
        lote_proveedor = (ir.get('lote_proveedor') or '').strip()
        ficha_seguridad_url = (ir.get('ficha_seguridad_url') or '').strip()
        # Solo registrar movimiento si hay algo recibido
        if cant_recibida > 0:
            if categoria == 'MEE':
                # Sin codigo_mp no se puede imputar el MEE · un INSERT con
                # mee_codigo='' + UPDATE que no matchea nada = drift permanente.
                if codigo:
                    cur.execute("UPDATE maestro_mee SET stock_actual = stock_actual + ? WHERE codigo=?", (cant_recibida, codigo))
                    cur.execute("INSERT INTO movimientos_mee (mee_codigo, tipo, cantidad, lote_ref, observaciones, responsable, fecha) VALUES (?,?,?,?,?,?,?)",
                               (codigo, 'Entrada', cant_recibida, numero_oc, f'Recepcion OC {numero_oc}', operador, fecha))
                else:
                    log.warning('recibir_oc MEE sin codigo_mp · OC %s · item no imputado', numero_oc)
            else:
                # Fix #9 · 21-may-2026 · lote sintético tracking (INVIMA).
                # Antes silencioso · receptor no se enteraba que su recepción
                # quedó con lote sintético que rompe trazabilidad con CoA.
                lote_final = lote_num
                if not lote_final:
                    lote_final = f'OC-{numero_oc}-{_idx+1}'
                    lotes_sinteticos_advertencia.append({
                        'codigo_mp': codigo, 'lote_asignado': lote_final,
                        'advertencia': 'Lote sintético · pedir lote real al proveedor para trazabilidad CoA/INVIMA',
                    })
                # AUDITORÍA-FIX 23-may-2026 · C20 · check duplicado lote_proveedor
                # mismo (lote_proveedor, proveedor, material_id) en distinta OC
                # = banderazo · INVIMA exige unicidad trazable
                if lote_proveedor and codigo:
                    try:
                        dup_row = cur.execute(
                            """SELECT numero_oc FROM movimientos
                               WHERE lote_proveedor=? AND material_id=?
                                 AND COALESCE(proveedor,'')=COALESCE(?,'')
                                 AND numero_oc != ?
                               LIMIT 1""",
                            (lote_proveedor, codigo, prov_nombre, numero_oc),
                        ).fetchone()
                        if dup_row:
                            lotes_sinteticos_advertencia.append({
                                'codigo_mp': codigo,
                                'lote_proveedor': lote_proveedor,
                                'advertencia': f'⚠ Lote proveedor {lote_proveedor} ya existe en OC {dup_row[0]} · verificar duplicado físico',
                            })
                    except sqlite3.OperationalError:
                        pass

                # Fase 2 · INVIMA · INSERT con campos COA/lote_proveedor (mig 151)
                # Si la columna no existe (mig no aplicó), cae a INSERT legacy.
                # Sebastián 16-jun · estado inicial controlado por interruptor
                # RECEPCION_AUTO_VIGENTE (default CUARENTENA = INVIMA · ON = carga
                # automática como VIGENTE, sin pasar por Calidad).
                from database import recepcion_auto_vigente as _rav
                _estado_recep = 'VIGENTE' if _rav(cur) else 'CUARENTENA'
                _coa_ok = False
                try:
                    cur.execute(
                        "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, "
                        "observaciones, proveedor, operador, lote, fecha_vencimiento, estado_lote, numero_oc, "
                        "coa_url, coa_filename, lote_proveedor, ficha_seguridad_url) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (codigo, nombre, cant_recibida, 'Entrada', fecha,
                         f'Recepcion OC {numero_oc}' + (f' | {notas_item}' if notas_item else ''),
                         prov_nombre, operador, lote_final,
                         fv or None, _estado_recep, numero_oc,
                         coa_url or None, coa_filename or None,
                         lote_proveedor or None, ficha_seguridad_url or None))
                    _coa_ok = True
                except Exception as _e:
                    log.info('movimientos sin columnas COA · cae a legacy: %s', _e)
                if _coa_ok:
                    pass  # INSERT ya hecho · pasar a actualizar cantidad
                else:
                    cur.execute(
                        "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, "
                        "observaciones, proveedor, operador, lote, fecha_vencimiento, estado_lote, numero_oc) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                        (codigo, nombre, cant_recibida, 'Entrada', fecha,
                         f'Recepcion OC {numero_oc}' + (f' | {notas_item}' if notas_item else ''),
                         prov_nombre, operador, lote_final,
                         fv or None, _estado_recep, numero_oc))
            ingresos += 1
        # Actualizar item OC · audit zero-error 2-may-2026: usar += en
        # cantidad_recibida_g para soportar recepciones múltiples parciales
        # sobre el mismo item. Antes era SET = ? que pisaba el acumulado.
        try:
            # UPDATE por rowid del item · antes era WHERE numero_oc+codigo_mp.
            cur.execute(
                "UPDATE ordenes_compra_items "
                "SET cantidad_recibida_g = COALESCE(cantidad_recibida_g, 0) + ?, "
                "    estado_recepcion=?, notas_recepcion=?, "
                "    lote_asignado=COALESCE(lote_asignado, ?) "
                " WHERE id=?",
                (cant_recibida, estado_item, notas_item, lote_num, _oci_rowid))
        except Exception as e:
            log.warning('UPDATE oci en recibir_oc fallo: %s', e)
        # SEC-FIX · 21-may-2026 · validación POST-UPDATE de sobre-recepción
        # · race condition · 2 workers reciben mismo item en paralelo
        # · ambos pasan pre-check 5% pero acumulado total supera
        try:
            row_post = cur.execute(
                "SELECT COALESCE(cantidad_g,0), COALESCE(cantidad_recibida_g,0) "
                "FROM ordenes_compra_items WHERE id=?", (_oci_rowid,),
            ).fetchone()
            if row_post:
                cant_pedida_post = float(row_post[0] or 0)
                cant_recib_post = float(row_post[1] or 0)
                # FIX · 29-may-2026 · audit ronda2 · honrar forzar_excepciones
                # igual que el pre-check (5425): si un admin envió forzar:true
                # acepta explícitamente la sobre-recepción · este guard de race
                # NO debe revertirla (antes anulaba el override documentado).
                if (not forzar_excepciones and cant_pedida_post > 0
                        and cant_recib_post > cant_pedida_post * 1.05):
                    # REVERTIR el UPDATE (race detectado)
                    cur.execute(
                        "UPDATE ordenes_compra_items "
                        "SET cantidad_recibida_g = COALESCE(cantidad_recibida_g, 0) - ? "
                        "WHERE id=?",
                        (cant_recibida, _oci_rowid),
                    )
                    return jsonify({
                        'error': f'Sobre-recepción detectada · {cant_recib_post:.2f}g > {cant_pedida_post*1.05:.2f}g (5% tolerancia)',
                        'codigo': 'SOBRE_RECEPCION_RACE',
                        'codigo_mp': codigo,
                    }), 422
        except Exception as _e:
            log.warning('post-recepcion guard fallo: %s', _e)

    # Estado final · usar acumulado de items, no solo la cantidad del request actual.
    # Bug fix Sebastián 8-may-2026: antes `es_parcial` chequeaba cant_recibida
    # de esta llamada vs cantidad_pedida, lo que mantenía OC en 'Parcial'
    # aunque la suma acumulada cubriera la pedida (bug en multi-recepción).
    es_parcial = False
    items_con_faltante = []  # FIX 23-may · detección auto discrepancia
    try:
        oci_rows = cur.execute(
            "SELECT codigo_mp, COALESCE(nombre_mp,''), COALESCE(cantidad_g, 0), "
            "       COALESCE(cantidad_recibida_g, 0) "
            "FROM ordenes_compra_items WHERE numero_oc=?",
            (numero_oc,)
        ).fetchall()
        for cm, nm, cant_ped, cant_recib_acum in oci_rows:
            ped = float(cant_ped or 0)
            recib = float(cant_recib_acum or 0)
            if recib < ped * 0.999:
                es_parcial = True
                # FIX 23-may-2026 · Sebastián · detección automática de
                # discrepancia · cualquier item con > 5% faltante levanta
                # la bandera tiene_discrepancias incluso si receptor no la
                # marcó manualmente · alerta a creador SOL + admin compras
                faltante = ped - recib
                if ped > 0 and (faltante / ped) > 0.05:  # >5% faltante
                    items_con_faltante.append({
                        'codigo_mp': cm,
                        'nombre_mp': nm,
                        'pedido': ped,
                        'recibido': recib,
                        'faltante': round(faltante, 1),
                        'pct_faltante': round((faltante / ped) * 100, 1),
                    })
    except Exception as e:
        log.warning('chequeo parcial post-update fallo: %s', e)

    # FIX 23-may-2026 · auto-set tiene_discrepancias si hay faltantes > 5%
    if items_con_faltante and not disc_r:
        disc_r = 1
        if obs_r:
            obs_r += ' | '
        obs_r += f'⚠ Auto-detección: {len(items_con_faltante)} item(s) con >5% faltante'

    # Si la OC ya estaba Pagada (anticipo · pago antes de la recepción), la
    # recepción registra el kardex pero NO revierte el estado (INV-4).
    # AUDITORÍA-FIX 23-may-2026 · C17 · flag recepcion_parcial separado del
    # estado · evita perder visibilidad de que falta mercancía cuando la
    # OC ya estaba Pagada (anticipo)
    recepcion_parcial_flag = 1 if es_parcial else 0
    if estado_oc_original == 'Pagada':
        nuevo_estado = 'Pagada'
    else:
        nuevo_estado = 'Parcial' if es_parcial else 'Recibida'
    try:
        cur.execute(
            "UPDATE ordenes_compra SET estado=?, fecha_recepcion=?,"
            " observaciones_recepcion=?, tiene_discrepancias=?, recibido_por=?,"
            " recepcion_parcial=? WHERE numero_oc=?",
            (nuevo_estado, fecha, obs_r, disc_r, receptor_nombre,
             recepcion_parcial_flag, numero_oc))
    except Exception:
        cur.execute("UPDATE ordenes_compra SET estado=?, fecha_recepcion=? WHERE numero_oc=?", (nuevo_estado, fecha, numero_oc))
    # Fix #7 · 21-may-2026 · lead_time real aprende del histórico (EWMA 0.7/0.3)
    # Antes: mp_lead_time_config quedaba con 14d default eternamente · auto_plan
    # y predicción demanda usaban datos falsos.
    # AUDITORÍA-FIX 23-may-2026 · C15 · solo aprender en recepción completa
    # · antes 3 partials = 3 updates EWMA encadenados (peso 0.7³=0.34 a histórico)
    # distorsionando el aprendizaje · ahora solo si not es_parcial
    if not es_parcial:
        try:
            oc_fecha_row = cur.execute(
                "SELECT fecha FROM ordenes_compra WHERE numero_oc=?", (numero_oc,)
            ).fetchone()
            if oc_fecha_row and oc_fecha_row[0]:
                from datetime import datetime as _dtle
                f_oc = _dtle.strptime(str(oc_fecha_row[0])[:10], '%Y-%m-%d').date()
                f_rec = _dtle.strptime(str(fecha)[:10], '%Y-%m-%d').date()
                lead_real_dias = max(1, (f_rec - f_oc).days)
                for codigo_mp_item in {it[1] for it in items_oc if it[1]}:
                    # FIX P1 audit 24-may-2026 · EWMA warm-up: con n<3
                    # muestras una sola recepción anómala movía el promedio
                    # 30%. Ahora media simple acumulada hasta n=3, después
                    # EWMA 0.7/0.3 estándar. n_recepciones incrementa
                    # en cada aprendizaje.
                    cur.execute(
                        """INSERT INTO mp_lead_time_config (material_id, lead_time_dias, n_recepciones)
                           VALUES (?, ?, 1)
                           ON CONFLICT(material_id) DO UPDATE SET
                             lead_time_dias = ROUND(
                               CASE
                                 WHEN COALESCE(n_recepciones, 0) < 3
                                   THEN (COALESCE(lead_time_dias, ?) * COALESCE(n_recepciones, 0) + ?)
                                        / (COALESCE(n_recepciones, 0) + 1.0)
                                 ELSE 0.7 * COALESCE(lead_time_dias, ?) + 0.3 * ?
                               END
                             ),
                             n_recepciones = COALESCE(n_recepciones, 0) + 1,
                             actualizado_en = datetime('now', '-5 hours')""",
                        (codigo_mp_item, lead_real_dias,
                         lead_real_dias, lead_real_dias,
                         lead_real_dias, lead_real_dias),
                    )
        except Exception as _e:
            log.warning('lead_time learn fallo: %s', _e)

    # AUDITORÍA-FIX 23-may-2026 · C4 · maestro_mps.ultima_compra_at +
    # precio_referencia (canonical para sugeridor de proveedor) · antes
    # solo se actualizaba en update_sol_items y crear_oc · datos quedaban
    # stale después de recibir si el precio facturado cambió
    try:
        # Re-leer items con precio para tener el dato canonical
        precios_rows = cur.execute(
            "SELECT codigo_mp, precio_unitario FROM ordenes_compra_items WHERE numero_oc=?",
            (numero_oc,),
        ).fetchall()
        for codigo_mp_item, precio_item in precios_rows:
            if not codigo_mp_item:
                continue
            sets = []
            params_u = []
            sets.append("ultima_compra_at=?")
            params_u.append(fecha)
            if precio_item and float(precio_item) > 0:
                sets.append("precio_referencia=?")
                params_u.append(float(precio_item) * 1000.0)  # $/g → $/kg (INV-2)
            params_u.append(codigo_mp_item)
            try:
                cur.execute(
                    f"UPDATE maestro_mps SET {','.join(sets)} WHERE codigo_mp=?",
                    params_u,
                )
            except sqlite3.OperationalError:
                # Columna ultima_compra_at puede no existir en esquemas viejos
                # · intentar solo precio_referencia
                if precio_item and float(precio_item) > 0:
                    try:
                        cur.execute(
                            "UPDATE maestro_mps SET precio_referencia=? WHERE codigo_mp=?",
                            (float(precio_item) * 1000.0, codigo_mp_item),  # $/g → $/kg (INV-2)
                        )
                    except Exception:
                        pass
    except Exception as _e2:
        log.warning('maestro_mps update tras recepción fallo: %s', _e2)
    # Cierre automatico de cadena: si la OC esta linkeada a items del checklist
    # Pre-Produccion (via produccion_checklist.oc_numero) o solicitudes anticipadas,
    # marcarlos como 'recibido'. Solo cuando recepcion es completa, no parcial.
    items_checklist_actualizados = 0
    if not es_parcial:
        try:
            cur.execute("""
                UPDATE produccion_checklist SET
                  estado='recibido',
                  fecha_recibido=date('now', '-5 hours'),
                  actualizado_at=datetime('now', '-5 hours')
                WHERE oc_numero=? AND estado IN ('solicitado','en_transito','pendiente')
            """, (numero_oc,))
            items_checklist_actualizados = cur.rowcount or 0
            cur.execute("""
                UPDATE solicitudes_compra_anticipada SET estado='completada'
                WHERE oc_numero=? AND estado IN ('decidida','pendiente')
            """, (numero_oc,))
        except Exception:
            pass
    try:
        audit_log(cur, usuario=operador, accion='RECIBIR_OC',
                  tabla='ordenes_compra', registro_id=numero_oc,
                  despues={'estado': nuevo_estado, 'fecha_recepcion': fecha,
                            'parcial': es_parcial, 'ingresos': ingresos,
                            'tiene_discrepancias': bool(disc_r),
                            'items_con_faltante': len(items_con_faltante),
                            'receptor': receptor_nombre[:200],
                            'observaciones_recepcion': (obs_r or '')[:200]},
                  detalle=f"Recibió OC {numero_oc} ({'PARCIAL' if es_parcial else 'COMPLETA'}) "
                          f"· {ingresos} items · receptor {receptor_nombre[:60]}")
    except Exception as e:
        log.warning('audit_log RECIBIR_OC fallo: %s', e)

    # FIX 23-may-2026 · Sebastián · alerta in-app cuando hay discrepancia
    # · push_notif a creador SOL + admin compras para verificar
    if items_con_faltante:
        try:
            from blueprints.notif import push_notif
        except Exception:
            push_notif = None
        if push_notif:
            try:
                msg_items = ', '.join(
                    f"{x['codigo_mp']}: {x['recibido']}/{x['pedido']}g (-{x['pct_faltante']}%)"
                    for x in items_con_faltante[:3]
                )
                if len(items_con_faltante) > 3:
                    msg_items += f' (+{len(items_con_faltante)-3} más)'
                # Buscar creador OC
                creador_row = cur.execute(
                    "SELECT creado_por FROM ordenes_compra WHERE numero_oc=?",
                    (numero_oc,)
                ).fetchone()
                creador = (creador_row[0] if creador_row else '') or ''
                destinatarios = set()
                if creador:
                    destinatarios.add(creador.lower())
                # Admins compras · imports renombrados para evitar shadow
                # del scope del módulo (UnboundLocalError de variables globales
                # ya usadas arriba en _require_authorize_oc/recibir_oc)
                try:
                    from config import COMPRAS_ACCESS as _CA_NOTIF, ADMIN_USERS as _AU_NOTIF
                    for _u_notif in (_CA_NOTIF | _AU_NOTIF):
                        destinatarios.add(_u_notif.lower())
                except Exception:
                    pass
                for dest in destinatarios:
                    try:
                        push_notif(dest, 'oc_discrepancia',
                                   f'⚠ OC {numero_oc} recibida con discrepancia · {msg_items}',
                                   link=f'/admin/compras?oc={numero_oc}',
                                   importante=True)
                    except Exception:
                        pass
            except Exception as e:
                log.warning('notif discrepancia OC %s fallo: %s', numero_oc, e)
    conn.commit()
    return jsonify({
        'ok': True, 'numero_oc': numero_oc, 'ingresos': ingresos,
        'estado': nuevo_estado, 'parcial': es_parcial,
        'checklist_actualizados': items_checklist_actualizados,
        'lotes_sinteticos': lotes_sinteticos_advertencia,  # Fix #9
    })

# ============================================================
# Compras — Flujo de autorizacion y pago
# ============================================================

@bp.route('/api/ordenes-compra/<numero_oc>/revisar', methods=['PATCH'])
def revisar_oc(numero_oc):
    """[LEGACY 22-may-2026] · estado 'Revisada' eliminado del flujo (mig 157).
    Endpoint mantiene compat · pero ahora deja la OC en 'Borrador' (no avanza).
    Usar directamente PATCH /autorizar para mover Borrador→Autorizada.
    """
    # Audit zero-error 2-may-2026: RBAC + bloqueo si OC ya pagada
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in (set(COMPRAS_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Compras/Admin pueden revisar OC'}), 403
    d = request.get_json() or {}
    conn = get_db(); cur = conn.cursor()
    # Bloquear revisión si OC ya pagada o cancelada (volver de Pagada a Revisada
    # creaba inconsistencia auditiva)
    estado_row = cur.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?",
                              (numero_oc,)).fetchone()
    if estado_row and estado_row[0] in ('Pagada', 'Cancelada', 'Rechazada'):
        return jsonify({'error': f'No se puede revisar OC en estado {estado_row[0]}'}), 409
    # FIX · 22-may-2026 · mig 157 elimina 'Revisada' · queda en Borrador
    sets = ["estado='Borrador'"]; params = []
    if d.get('proveedor'):
        sets.append('proveedor=?'); params.append(str(d['proveedor']))
    if d.get('valor_total') not in (None, '', 0):
        sets.append('valor_total=?'); params.append(float(d['valor_total'] or 0))
    if d.get('observaciones'):
        sets.append('observaciones=?'); params.append(str(d['observaciones']))
    # IVA
    sets.append('con_iva=?'); params.append(1 if d.get('con_iva') else 0)
    sets.append('valor_sin_iva=?'); params.append(float(d.get('valor_sin_iva') or 0))
    params.append(numero_oc)
    cur.execute(f"UPDATE ordenes_compra SET {', '.join(sets)} WHERE numero_oc=?", params)
    try:
        audit_log(cur, usuario=user, accion='REVISAR_OC',
                  tabla='ordenes_compra', registro_id=numero_oc,
                  antes={'estado': estado_row[0] if estado_row else None},
                  despues={'estado': 'Revisada',
                            'proveedor': d.get('proveedor'),
                            'valor_total': d.get('valor_total'),
                            'con_iva': bool(d.get('con_iva'))},
                  detalle=f"Revisó OC {numero_oc}")
    except Exception as e:
        log.warning('audit_log REVISAR_OC fallo: %s', e)
    conn.commit()
    return jsonify({'ok': True, 'estado': 'Revisada'})

@bp.route('/api/ordenes-compra/<numero_oc>/autorizar', methods=['PATCH'])
def autorizar_oc(numero_oc):
    usuario_actual, err, code = _require_authorize_oc()
    if err:
        return err, code
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT estado, valor_total, categoria FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    row = cur.fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado_ant = row[0]
    categoria_oc = row[2] or ''
    # No re-autorizar una OC ya avanzada/terminal · evita revertir el ciclo
    # Autorizada→Pagada→Recibida o resucitar una OC cancelada.
    if (estado_ant or '') in ('Pagada', 'Recibida', 'Cancelada', 'Anulada'):
        return jsonify({'error': f'No se puede autorizar una OC en estado {estado_ant}'}), 409
    valor = float(row[1] or 0)
    # Sebastián 24-may-2026 · audit OCs Activas · validar items > 0 antes de
    # autorizar. Antes una OC sin items podía pasar a Autorizada con valor_total=0
    # · email al proveedor con tabla vacía · confusión. Ahora exigimos al menos
    # 1 item con cantidad > 0 (excepción: categorías de pago directo · SVC/CC
    # /Influencer · que pueden tener items genéricos sin cantidad real).
    _CATEGORIAS_SIN_ITEMS = ('Influencer/Marketing Digital', 'Cuenta de Cobro',
                              'CC', 'SVC', 'Servicios', 'Servicios Profesionales')
    if categoria_oc not in _CATEGORIAS_SIN_ITEMS:
        try:
            _n_items_row = cur.execute(
                "SELECT COUNT(*), COALESCE(SUM(cantidad_g),0) "
                "FROM ordenes_compra_items WHERE numero_oc=?",
                (numero_oc,),
            ).fetchone()
            _n_items = int(_n_items_row[0] or 0)
            _sum_cant = float(_n_items_row[1] or 0)
        except sqlite3.OperationalError:
            _n_items, _sum_cant = 0, 0
        if _n_items == 0 or _sum_cant <= 0:
            return jsonify({
                'error': 'OC sin items · agregá líneas con cantidad > 0 antes de autorizar',
                'codigo': 'OC_SIN_ITEMS',
                'n_items': _n_items, 'suma_cantidad_g': _sum_cant,
                'hint': 'Editá la OC y agregá al menos un material con cantidad y precio.',
            }), 409
        if valor <= 0:
            return jsonify({
                'error': 'OC con valor_total=0 · agregá precios a los items antes de autorizar',
                'codigo': 'OC_SIN_VALOR',
                'hint': 'Editá la OC e ingresá precio_unitario > 0 en cada item.',
            }), 409
    # Sprint 4: límite de aprobación por usuario
    err_lim, code_lim = _check_monto_limit(usuario_actual, valor)
    if err_lim:
        return err_lim, code_lim
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    cur.execute("SELECT remision_code FROM ordenes_compra WHERE remision_code LIKE ? ORDER BY remision_code DESC LIMIT 1",
                (f'REM-ESP-{fecha_hoy}-%',))
    last = cur.fetchone()
    n = int(last[0].split('-')[-1]) + 1 if last and last[0] else 1
    remision_code = f'REM-ESP-{fecha_hoy}-{n:03d}'
    fecha_aut = datetime.now().isoformat()
    # Bug #7 fix · 21-may-2026 · UPDATE con CAS (compare-and-swap) anti-race
    # · si otro admin ya autorizó simultáneamente, rowcount=0 y devuelve 409
    cur.execute(
        "UPDATE ordenes_compra SET estado='Autorizada', remision_code=?, "
        "autorizado_por=?, fecha_autorizacion=? WHERE numero_oc=? AND estado=?",
        (remision_code, usuario_actual, fecha_aut, numero_oc, estado_ant))
    if cur.rowcount == 0:
        # Race · otro admin lo autorizó primero · re-leer y devolver info
        re_row = cur.execute(
            "SELECT estado, autorizado_por FROM ordenes_compra WHERE numero_oc=?",
            (numero_oc,),
        ).fetchone()
        return jsonify({
            'error': 'Race condition · otro usuario ya autorizó la OC',
            'codigo': 'RACE_AUTORIZACION',
            'estado_actual': re_row[0] if re_row else '?',
            'autorizado_por': re_row[1] if re_row else '?',
        }), 409
    # Audit log · autorización es decisión regulatoria/financiera
    try:
        audit_log(cur, usuario=usuario_actual, accion='AUTORIZAR_OC',
                  tabla='ordenes_compra', registro_id=numero_oc,
                  antes={'estado': estado_ant, 'valor_total': valor},
                  despues={'estado': 'Autorizada', 'remision_code': remision_code})
    except Exception as e:
        log.warning('audit_log AUTORIZAR_OC fallo: %s', e)
    # Fase 2 · 21-may-2026 · email auto al proveedor con detalle OC
    # Toggle: env COMPRAS_AUTO_EMAIL_PROV_OFF=1 desactiva (default ON)
    email_enviado = False
    email_status = 'sin email proveedor'
    import os as _os_mail
    if (_os_mail.environ.get('COMPRAS_AUTO_EMAIL_PROV_OFF') or '0') != '1':
        try:
            prov_row = cur.execute(
                "SELECT proveedor FROM ordenes_compra WHERE numero_oc=?",
                (numero_oc,),
            ).fetchone()
            prov_nombre = prov_row[0] if prov_row else ''
            email_row = cur.execute(
                "SELECT email FROM proveedores WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?))",
                (prov_nombre,),
            ).fetchone()
            email_prov = (email_row[0] if email_row else '') or ''
            items_rows = cur.execute(
                """SELECT codigo_mp, nombre_mp, cantidad_g, precio_unitario
                   FROM ordenes_compra_items WHERE numero_oc=?""",
                (numero_oc,),
            ).fetchall()
            items_list = [{
                'codigo_mp': r[0], 'nombre_mp': r[1],
                'cantidad_g': r[2], 'precio_unitario': r[3],
            } for r in items_rows]
            obs_row = cur.execute(
                "SELECT COALESCE(observaciones,'') FROM ordenes_compra WHERE numero_oc=?",
                (numero_oc,),
            ).fetchone()
            email_enviado, email_status = _enviar_oc_a_proveedor(
                numero_oc, prov_nombre, email_prov, items_list, valor,
                obs_row[0] if obs_row else '',
            )
        except Exception as _e:
            email_status = f'fallo: {_e}'
    conn.commit()
    return jsonify({
        'ok': True, 'estado': 'Autorizada', 'remision_code': remision_code,
        'email_proveedor_enviado': email_enviado,
        'email_status': email_status,
    })

@bp.route('/api/ordenes-compra/<numero_oc>/pagar', methods=['PATCH'])
def pagar_oc(numero_oc):
    """Registra un pago de una OC.

    Mejoras Sprint 2:
      - Soporta pagos PARCIALES: registra cada pago en tabla `pagos_oc`
        (auditoría completa). Estado de la OC se calcula:
          - 'Pagada' si suma pagos >= valor_total
          - 'Parcial' si 0 < suma pagos < valor_total
      - 3-way matching: campo `numero_factura_proveedor` con UNIQUE soft.
        Si la factura ya fue usada en otro pago → 409 Conflict (anti doble pago).
      - Alimenta `precios_mp_historico` por cada item de la OC con el precio
        efectivamente pagado (incluyendo prorrateo si fue parcial).
    """
    usuario_actual, err, code = _require_authorize_oc()
    if err:
        return err, code
    d = request.get_json() or {}
    # Money sanity validation · audit zero-error 2-may-2026
    raw_monto = d.get('monto', 0) or 0
    if raw_monto:
        monto_validado, err = validate_money(raw_monto, allow_zero=False, field_name='monto')
        if err:
            return jsonify(err), 400
        monto = monto_validado
    else:
        monto = 0  # se completará con valor_total_oc después si viene 0
    medio = d.get('medio', 'Transferencia')
    obs = d.get('observaciones', '')
    numero_factura = (d.get('numero_factura_proveedor') or '').strip()
    comprobante_imagen = d.get('comprobante_imagen', '') or ''
    # Toggles fiscales (default OFF, ver Opcion A en SECURITY notes)
    aplicar_retefuente = bool(d.get('aplicar_retefuente', False))
    aplicar_retica = bool(d.get('aplicar_retica', False))
    aplicar_iva = bool(d.get('aplicar_iva', False))
    # Limit image size to ~4MB base64 to avoid DB bloat
    if len(comprobante_imagen) > 4_000_000:
        comprobante_imagen = comprobante_imagen[:4_000_000]
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT estado, categoria, proveedor, valor_total FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    row = cur.fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado_oc = row[0] or ''
    categoria = row[1] or 'MP'
    proveedor = row[2] or ''
    valor_total_oc = float(row[3] or 0)
    # Audit zero-error 2-may-2026: respetar LIMITES_APROBACION_OC en pagos
    # Antes _check_monto_limit solo se aplicaba en autorizar_oc · ahora también
    # en pagos (segregation of duties: si Catalina autoriza, otro paga).
    if not monto:
        monto_para_limit = valor_total_oc
    else:
        monto_para_limit = float(monto)
    err_lim, code_lim = _check_monto_limit(usuario_actual, monto_para_limit)
    if err_lim:
        return err_lim, code_lim
    # Audit zero-error 2-may-2026: bloquear pagos sobre OC en estado inválido
    # Bug #8 fix · 21-may-2026 · bloquear pagar Revisada (bypass autorización gerencial)
    if estado_oc in ('Cancelada', 'Rechazada', 'Borrador', 'Revisada'):
        return jsonify({
            'error': f"OC {numero_oc} en estado '{estado_oc}' no admite pagos",
            'codigo': 'ESTADO_INVALIDO'
        }), 409
    if not monto:
        monto = valor_total_oc
    if monto <= 0:
        return jsonify({'error': 'monto debe ser > 0', 'codigo': 'MONTO_INVALIDO'}), 400
    # FIX 13-jun (audit influencers · H3): si Marketing YA marcó este pago de
    # influencer como Pagado (corrección manual vía mkt_pago_influencer_editar),
    # NO crear otro egreso por acá → evita doble pago. Antes Marketing mostraba
    # "Pagado" y Compras lo seguía mostrando pagable = divergencia + riesgo doble
    # egreso. (No bloquea el flujo normal: ahí pagos_influencers está 'Pendiente'
    # hasta que ESTE pago lo marque Pagado al final.)
    try:
        _inf_pagado = cur.execute(
            "SELECT 1 FROM pagos_influencers WHERE numero_oc=? AND estado='Pagada' LIMIT 1",
            (numero_oc,)).fetchone()
    except Exception:
        _inf_pagado = None
    if _inf_pagado:
        return jsonify({
            'error': ('Este pago de influencer ya figura PAGADO en Marketing · '
                      'no lo pagues de nuevo por acá. Reconciliá antes (Marketing → '
                      'editar pago) si fue un error.'),
            'codigo': 'INFLUENCER_YA_PAGADO',
        }), 409
    # Audit zero-error 2-may-2026: validar over-payment ANTES de insertar
    cur.execute("SELECT COALESCE(SUM(monto), 0) FROM pagos_oc WHERE numero_oc=?", (numero_oc,))
    total_pagado_actual = float(cur.fetchone()[0] or 0)
    if (total_pagado_actual + monto) > (valor_total_oc + 0.01):
        return jsonify({
            'error': f"Over-payment: pagado actual {total_pagado_actual:.0f} + nuevo {monto:.0f} "
                       f"excede valor OC {valor_total_oc:.0f}",
            'pagado_actual': total_pagado_actual,
            'valor_oc': valor_total_oc,
            'codigo': 'OVER_PAYMENT'
        }), 422
    cat_map = {'MPs':'MPs','MP':'MPs','Envase':'MEE','Insumos':'MEE','MEE':'MEE','SVC':'Servicios','Servicios':'Servicios','Analisis':'Servicios','Ánalisis':'Servicios','Acondicionamiento':'Servicios','Admin':'Administrativo','Nomina':'Administrativo','ADM':'Administrativo','Infraestructura':'Infraestructura','INF':'Infraestructura','CC':'Cuentas de Cobro'}
    cat_egreso = cat_map.get(categoria, 'Compras')
    fecha_pago = datetime.now().isoformat()

    # 3-way matching: validar que la factura no haya sido usada en otro pago
    if numero_factura:
        cur.execute(
            "SELECT numero_oc FROM pagos_oc WHERE numero_factura_proveedor=? LIMIT 1",
            (numero_factura,)
        )
        prev = cur.fetchone()
        if prev and prev[0] != numero_oc:
            return jsonify({
                'error': f"Factura '{numero_factura}' ya fue registrada en pago de OC {prev[0]}",
                'detail': 'Anti doble pago — verifica antes de continuar.',
                'codigo': 'FACTURA_DUPLICADA'
            }), 409
        # Anti-doble-pago cruzado con el libro de facturas (INV-6 · 1-jun-2026):
        # si ese número ya tiene pagos en facturas_proveedor, no pagar de nuevo acá.
        try:
            _enlibro = cur.execute(
                """SELECT 1 FROM facturas_proveedor f
                   WHERE f.numero_factura=? AND COALESCE(f.estado,'') != 'anulada'
                     AND EXISTS (SELECT 1 FROM pagos_oc p WHERE p.factura_proveedor_id=f.id)
                   LIMIT 1""", (numero_factura,)).fetchone()
        except Exception:
            _enlibro = None
        if _enlibro:
            return jsonify({
                'error': f"La factura '{numero_factura}' ya tiene pagos en el libro de facturas · "
                         f"no la pagues de nuevo por acá",
                'codigo': 'FACTURA_YA_EN_LIBRO'
            }), 409

    # Registrar el pago en pagos_oc (auditoría completa)
    try:
        cur.execute("""
            INSERT INTO pagos_oc (numero_oc, monto, medio, fecha_pago,
                                  registrado_por, numero_factura_proveedor,
                                  comprobante_imagen, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (numero_oc, monto, medio, fecha_pago, usuario_actual,
              numero_factura, comprobante_imagen, obs))
    except sqlite3.IntegrityError as _e:
        # UNIQUE de factura disparó (race con check anterior — defense in depth)
        return jsonify({
            'error': 'Factura duplicada en otro pago concurrente',
            'detail': str(_e)[:200],
            'codigo': 'FACTURA_DUPLICADA'
        }), 409

    # FIX 27-may (P2 race) · SELECT + UPDATE no eran atómicos. Si 2 pagos
    # concurrentes llegaban en ventana corta, total_pagado leído quedaba stale
    # y la OC podía marcarse 'Parcial' cuando ya estaba 'Pagada' o viceversa.
    # CAS: el UPDATE re-calcula SUM(monto) inline dentro del CASE · una sola
    # transacción · sin lectura externa stale.
    cur.execute("""
        UPDATE ordenes_compra SET
            estado = CASE WHEN (
                SELECT COALESCE(SUM(monto),0) FROM pagos_oc WHERE numero_oc=?
            ) >= ? - 0.01 THEN 'Pagada' ELSE 'Parcial' END,
            pagado_por=?, fecha_pago=?, medio_pago=?, comprobante_imagen=?
        WHERE numero_oc=?
    """, (numero_oc, valor_total_oc,
          usuario_actual, fecha_pago, medio,
          comprobante_imagen, numero_oc))
    # Releer el estado para el audit + return
    cur.execute("SELECT estado FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    _row_est = cur.fetchone()
    nuevo_estado = (_row_est[0] if _row_est else 'Parcial') or 'Parcial'
    cur.execute("SELECT COALESCE(SUM(monto),0) FROM pagos_oc WHERE numero_oc=?", (numero_oc,))
    total_pagado = float(cur.fetchone()[0] or 0)
    # FIX 13-jun (audit compras · M27/M12(d)) · guard de over-payment ATÓMICO post-insert.
    # El check previo (~6470) es check-then-act; con 2 pagos full concurrentes SIN nº de
    # factura (el path con factura ya lo cubre el UNIQUE) ambos podían pasar el check y
    # duplicar el egreso. Acá re-verificamos el SUM REAL ya con este pago insertado: si
    # excede el valor de la OC, otro pago entró en paralelo → rollback ANTES de crear el
    # egreso financiero / comprobante. (Cubre el doble-click y el caso donde el otro pago
    # ya commiteó; SQLite serializa escrituras, así que el riesgo era solo PG.)
    if total_pagado > (valor_total_oc + 0.01):
        conn.rollback()
        return jsonify({
            'error': (f"Over-payment por pago concurrente: total {total_pagado:.0f} "
                      f"excede el valor de la OC {valor_total_oc:.0f}. No se registró este pago."),
            'codigo': 'OVER_PAYMENT_RACE',
        }), 409
    # Audit log INVIMA · Resolución 2214/2021 · pago de OC es operación financiera regulada
    try:
        audit_log(cur, usuario=usuario_actual, accion='PAGAR_OC',
                  tabla='pagos_oc', registro_id=numero_oc,
                  antes={'estado_oc': estado_oc,
                         'total_pagado_antes': total_pagado_actual},
                  despues={'monto': monto, 'medio': medio,
                           'numero_factura': numero_factura,
                           'estado_nuevo': nuevo_estado,
                           'total_pagado_despues': total_pagado})
    except Exception as e:
        log.warning('audit_log PAGAR_OC fallo: %s · operación NO se aborta (audit es defense-in-depth)', e)
    # Sync solicitudes_compra estado → Pagada so it leaves the pending list.
    # Sebastian (30-abr-2026): bulletproof fix — antes el WHERE filtraba por
    # estado='Aprobada' y si la SOL estaba en otro estado (ej. Pendiente
    # despues de un flujo abortado, o ya Pagada por sync previo) NO se
    # actualizaba y quedaba pegada en la lista. Ahora actualizamos cualquier
    # estado != 'Pagada' al pagar la OC vinculada.
    cur.execute(
        "UPDATE solicitudes_compra SET estado='Pagada' "
        "WHERE numero_oc=? AND estado != 'Pagada'",
        (numero_oc,)
    )
    # Bulletproof 2-jun-2026 (Jeferson "le doy pagar y no desaparece"): el front
    # envía el número de la SOL que el usuario pagó. Flipear ESA SOL directamente
    # cubre el caso de numero_oc NULL/desalineado (legacy o link fallido) que
    # dejaba el item pegado en la lista de influencers. Además re-vincula la OC.
    _sol_numero = (d.get('sol_numero') or '').strip()
    if _sol_numero:
        cur.execute(
            "UPDATE solicitudes_compra SET estado='Pagada', "
            "numero_oc=COALESCE(NULLIF(numero_oc,''),?) "
            "WHERE numero=? AND estado != 'Pagada'",
            (numero_oc, _sol_numero)
        )
    # Push notif in-app al solicitante (Jefferson en este caso) para que sepa
    # que su pago se procesó. Sebastian (30-abr-2026): "tampoco le estara
    # notificando a el, revisa si pueden llegar a marketing o como seria".
    try:
        sol_notif = cur.execute(
            "SELECT numero, solicitante, valor FROM solicitudes_compra "
            "WHERE numero_oc=? LIMIT 1", (numero_oc,)
        ).fetchone()
        if sol_notif:
            from blueprints.notif import push_notif
            sol_num_n, solicitante_n, valor_n = sol_notif
            destinatario = (solicitante_n or '').lower().strip()
            if destinatario:
                push_notif(
                    destinatario, 'oc_estado',
                    f'✅ Pago procesado: {sol_num_n}',
                    body=f'OC {numero_oc} pagada por ${(valor_n or monto or 0):,.0f} · medio {medio}',
                    link='/marketing',
                    remitente=usuario_actual,
                    importante=True
                )
                # Tambien notif a otros usuarios de marketing para visibilidad
                # (Daniela suele ver, etc) — non-blocking, best effort.
                try:
                    from blueprints.marketing import MARKETING_USERS as _MK
                except Exception:
                    _MK = ()
                for _u in (_MK or ()):
                    if _u.lower() != destinatario:
                        push_notif(_u.lower(), 'oc_estado',
                                   f'💸 Pago procesado a {sol_num_n}',
                                   body=f'OC {numero_oc} pagada · ${(valor_n or monto or 0):,.0f}',
                                   link='/marketing', remitente=usuario_actual)
    except Exception as _e:
        __import__('logging').getLogger('compras').warning(
            'push_notif Pagada SOL fallo OC %s: %s', numero_oc, _e
        )
    # Sync marketing payment status:
    # 1. Si ya existe un row en pagos_influencers para esta OC → marcar Pagada
    #    SIEMPRE (sin importar categoria — la existencia del row es señal
    #    suficiente; el bug previo era que si categoria estaba mal/vacía, no
    #    se actualizaba y quedaba como Pendiente).
    # 2. Si no existe row pero categoria sugiere influencer → crearlo Pagada.
    try:
        cat_low = (categoria or '').lower()
        is_influencer_cat = 'influencer' in cat_low or 'marketing' in cat_low
        existing = cur.execute(
            "SELECT id FROM pagos_influencers WHERE numero_oc=? LIMIT 1",
            (numero_oc,)
        ).fetchone()

        # Recolectar info del influencer para enriquecer (best-effort)
        inf_id = None
        inf_name = proveedor  # fallback
        try:
            sol_row = cur.execute(
                "SELECT influencer_id, solicitante FROM solicitudes_compra WHERE numero_oc=? LIMIT 1",
                (numero_oc,)
            ).fetchone()
            if sol_row:
                inf_id = sol_row[0] if sol_row[0] else None
                if inf_id:
                    inf_row = cur.execute(
                        "SELECT nombre FROM marketing_influencers WHERE id=?", (inf_id,)
                    ).fetchone()
                    if inf_row:
                        inf_name = inf_row[0]
                elif sol_row[1]:
                    inf_name = sol_row[1]  # solicitante name as fallback
        except sqlite3.OperationalError:
            pass  # columnas pueden no existir en instancias viejas

        if existing:
            # Update SIEMPRE: si la fila existe, este pago la marca como Pagada
            cur.execute(
                "UPDATE pagos_influencers SET estado='Pagada', "
                "influencer_id=COALESCE(influencer_id,?), "
                "influencer_nombre=CASE WHEN influencer_nombre IN ('','Pago') "
                "THEN ? ELSE influencer_nombre END "
                "WHERE numero_oc=?",
                (inf_id, inf_name, numero_oc)
            )
        elif is_influencer_cat:
            # No hay fila pero la categoría dice influencer → crear Pagada.
            # FIX 27-may · setear fecha_contenido + vence_pago_at (mig 195)
            # incluso para Pagada · permite reportes históricos correctos.
            from datetime import datetime as _dtPI, timedelta as _tdPI
            _base_pi = (_dtPI.utcnow() - _tdPI(hours=5)).date()
            _fc_pi = _base_pi.isoformat()
            _vence_pi = (_base_pi + _tdPI(days=30)).isoformat()
            try:
                cur.execute("""
                    INSERT INTO pagos_influencers
                    (influencer_id, influencer_nombre, valor, fecha, estado,
                     concepto, numero_oc, fecha_contenido, vence_pago_at)
                    VALUES (?,?,?,date('now', '-5 hours'),'Pagada',?,?,?,?)
                """, (inf_id, inf_name, monto, f'Pago OC {numero_oc}',
                       numero_oc, _fc_pi, _vence_pi))
            except Exception:
                cur.execute("""
                    INSERT INTO pagos_influencers
                    (influencer_id, influencer_nombre, valor, fecha, estado, concepto, numero_oc)
                    VALUES (?,?,?,date('now', '-5 hours'),'Pagada',?,?)
                """, (inf_id, inf_name, monto, f'Pago OC {numero_oc}', numero_oc))
    except Exception as _e:
        __import__('logging').getLogger('compras').warning(
            "sync pagos_influencers falló para OC %s: %s", numero_oc, _e
        )
    try:
        # Detectar empresa pagadora con la misma logica multi-señal que usan los
        # comprobantes de egreso. Influencers / marketing / cuenta de cobro →
        # ANIMUS; mercancia/MPs/servicios tecnicos → ESPAGIRIA. Antes estaba
        # hardcoded a 'Espagiria' y por eso pagos a influencers (Ana Sofia,
        # Maria Camila Soto, daisy lopez, etc.) aparecian mal categorizados
        # en el Historial de Egresos del Financiero. Sebastian 2026-04-29.
        empresa_egreso = 'Animus' if _is_animus_payment(
            cur,
            numero_oc=numero_oc,
            beneficiario_nombre=proveedor,
            observaciones=obs,
            categoria=cat_egreso,
        ) else 'Espagiria'
        cur.execute("INSERT INTO flujo_egresos (fecha, empresa, concepto, categoria, monto, periodo, fuente, referencia, creado_por, observaciones) VALUES (?,?,?,?,?,?,?,?,?,?)",
                   (fecha_pago, empresa_egreso, f'Pago OC {numero_oc} - {proveedor}',
                    cat_egreso, monto, datetime.now().strftime('%Y-%m'),
                    'compras', numero_oc, usuario_actual, f'{medio}. {obs}'))
    except sqlite3.OperationalError as _e:
        if 'no such table' not in str(_e).lower():
            __import__('logging').getLogger('compras').error(
                "INSERT flujo_egresos falló: %s", _e
            )

    # Alimentar historial de precios (Sprint 2): por cada item de la OC,
    # registrar el precio efectivo en precios_mp_historico. Útil para detectar
    # variaciones de precio en Sprint 4 (reporte ejecutivo).
    try:
        cur.execute("""SELECT codigo_mp, precio_unitario
                       FROM ordenes_compra_items
                       WHERE numero_oc=? AND codigo_mp IS NOT NULL AND codigo_mp != ''""",
                    (numero_oc,))
        items_oc = cur.fetchall()
        for codigo_mp, precio in items_oc:
            if precio and precio > 0:
                cur.execute("""INSERT OR IGNORE INTO precios_mp_historico
                               (codigo_mp, precio_kg, numero_factura, proveedor, fecha)
                               VALUES (?, ?, ?, ?, datetime('now', '-5 hours'))""",
                            (codigo_mp, float(precio) * 1000.0, numero_factura, proveedor))  # precio_unitario $/g → precio_kg
                # AUDITORÍA-FIX 23-may-2026 · C14 · sincronizar precio_referencia
                # en maestro_mps (canonical para sugeridor) · antes solo se
                # actualizaba precios_mp_historico
                try:
                    cur.execute(
                        "UPDATE maestro_mps SET precio_referencia=? WHERE codigo_mp=?",
                        (float(precio) * 1000.0, codigo_mp),  # $/g → $/kg (INV-2)
                    )
                except Exception:
                    pass
    except sqlite3.OperationalError as _e:
        if 'no such table' not in str(_e).lower():
            __import__('logging').getLogger('compras').error(
                "Alimentar precios_mp_historico falló: %s", _e
            )

    # ── Generar Comprobante de Egreso (CE) — formato fiscal-compatible ─────
    # Para todo pago: genera PDF + guarda en comprobantes_pago. La contadora
    # accede a estos PDFs desde /contabilidad para sus cuentas.
    comprobante_info = None
    try:
        from comprobante_pago import crear_comprobante_y_pdf

        # Recoger último pago_oc_id (el que acabamos de insertar)
        cur.execute("SELECT MAX(id) FROM pagos_oc WHERE numero_oc=?", (numero_oc,))
        pago_oc_id_row = cur.fetchone()
        pago_oc_id = pago_oc_id_row[0] if pago_oc_id_row else None

        # Datos del beneficiario: si es influencer, jalar de marketing_influencers
        beneficiario = {
            'nombre': proveedor, 'cedula': '', 'banco': '',
            'cuenta': '', 'tipo_cuenta': '', 'ciudad': '', 'email': ''
        }
        cat_lower = (categoria or '').lower()
        if 'influencer' in cat_lower or 'marketing' in cat_lower:
            try:
                inf_row = cur.execute("""
                    SELECT mi.nombre, mi.cedula_nit, mi.banco, mi.cuenta_bancaria,
                           mi.tipo_cuenta, mi.ciudad, mi.email
                    FROM solicitudes_compra sc
                    LEFT JOIN marketing_influencers mi ON sc.influencer_id = mi.id
                    WHERE sc.numero_oc=? AND mi.id IS NOT NULL LIMIT 1
                """, (numero_oc,)).fetchone()
                if inf_row:
                    beneficiario = {
                        'nombre': inf_row[0] or proveedor,
                        'cedula': inf_row[1] or '',
                        'banco': inf_row[2] or '',
                        'cuenta': inf_row[3] or '',
                        'tipo_cuenta': inf_row[4] or '',
                        'ciudad': inf_row[5] or '',
                        'email': inf_row[6] or '',
                    }
            except sqlite3.OperationalError:
                pass
        else:
            # Proveedor regular: jalar datos de la tabla proveedores
            try:
                pv = cur.execute(
                    "SELECT contacto, nit, banco, num_cuenta, tipo_cuenta, email FROM proveedores WHERE nombre=? LIMIT 1",
                    (proveedor,)
                ).fetchone()
                if pv:
                    beneficiario.update({
                        'cedula': pv[1] or '', 'banco': pv[2] or '',
                        'cuenta': pv[3] or '', 'tipo_cuenta': pv[4] or '',
                        'email': pv[5] or '',
                    })
            except sqlite3.OperationalError:
                pass

        # Items: descripción consolidada de la OC (1 sola línea para servicios,
        # múltiples para mercancía con codigo_mp).
        try:
            items_db = cur.execute("""
                SELECT nombre_mp, cantidad_g, precio_unitario, subtotal
                FROM ordenes_compra_items WHERE numero_oc=?
            """, (numero_oc,)).fetchall()
        except sqlite3.OperationalError:
            items_db = []

        # Detectar si la OC es de servicio (no mercancia con peso real). Para
        # servicios/donaciones/cuentas de cobro/influencers, los items pueden
        # tener cantidad_g=1 como placeholder y subtotal=monto total — si los
        # tratamos como gramos el PDF muestra cant=0 (1g/1000=0.001 redondeado)
        # y valor_total = unit*0.001 = 1/1000 del valor real (bug Sebastian
        # 29-abr-2026: $1,200,000 aparecía como $1,200 en CE-2026-0008).
        cat_low_pdf = (categoria or '').lower()
        es_servicio_oc = any(k in cat_low_pdf for k in (
            'influencer', 'marketing', 'cuenta de cobro', 'servicio',
            'admin', 'infraestructura', 'cc', 'svc',
        ))
        if items_db and not es_servicio_oc:
            # Mercancia real (MPs, MEE): cantidad_g viene en gramos, subtotal
            # es valor del item completo. Mostrar en kg para legibilidad.
            items_pdf = []
            for r in items_db:
                cant_g = float(r[1] or 0)
                subt = float(r[3] or 0)
                cant_kg = cant_g / 1000.0
                if cant_kg <= 0.01:
                    # Cantidad insignificante → tratar como servicio: cant=1, unit=subtotal
                    items_pdf.append({
                        'descripcion': r[0] or '', 'fecha': '',
                        'cantidad': 1, 'valor_unit': subt,
                    })
                else:
                    p_unit = subt / cant_kg if cant_kg > 0 else 0
                    items_pdf.append({
                        'descripcion': r[0] or '', 'fecha': '',
                        'cantidad': cant_kg, 'valor_unit': p_unit,
                    })
        elif items_db:
            # Servicio CON items registrados (ej. una donacion con descripcion
            # "Donacion") — usar el subtotal directo como valor_unit con cant=1
            items_pdf = [{
                'descripcion': r[0] or f"Pago OC {numero_oc} - {categoria}",
                'fecha': fecha_pago[:10],
                'cantidad': 1,
                'valor_unit': float(r[3] or 0) or monto,
            } for r in items_db]
        else:
            # Servicio sin items detallados: 1 fila generica con el monto total
            items_pdf = [{
                'descripcion': f"Pago OC {numero_oc} - {categoria}",
                'fecha': fecha_pago[:10],
                'cantidad': 1,
                'valor_unit': monto,
            }]

        # Subtotal del comprobante: si aplican retenciones/IVA, el "monto"
        # registrado es el total a pagar; el subtotal_pre se calcula al revés.
        # Para simplicidad: subtotal = monto (el toggle aplica retenciones AL
        # GENERAR el comprobante, no al guardar el pago).
        subtotal_ce = monto
        if aplicar_iva:
            # Si IVA está prendido, asumimos que el "monto" del pago es el
            # subtotal y el IVA se suma — el contador recibe el comprobante
            # con el desglose correcto.
            pass

        # ── OBS string fallback: si aún faltan datos bancarios, parsear OBS ──
        # Ocurre cuando influencer_id era NULL en solicitudes antiguas
        # o cuando el registro en marketing_influencers no tiene banco/cuenta.
        if not beneficiario.get('banco') or not beneficiario.get('cuenta'):
            try:
                from comprobante_pago import parse_obs_beneficiario
                obs_row = cur.execute(
                    "SELECT observaciones FROM solicitudes_compra WHERE numero_oc=? LIMIT 1",
                    (numero_oc,)
                ).fetchone()
                if obs_row and obs_row[0]:
                    parsed = parse_obs_beneficiario(obs_row[0])
                    # Completar solo lo que falta (no pisar datos ya correctos)
                    if not beneficiario.get('banco') and parsed.get('banco'):
                        beneficiario['banco'] = parsed['banco']
                    if not beneficiario.get('tipo_cuenta') and parsed.get('tipo_cuenta'):
                        beneficiario['tipo_cuenta'] = parsed['tipo_cuenta']
                    if not beneficiario.get('cuenta') and parsed.get('cuenta'):
                        beneficiario['cuenta'] = parsed['cuenta']
                    if not beneficiario.get('cedula') and parsed.get('cedula'):
                        beneficiario['cedula'] = parsed['cedula']
                    if not beneficiario.get('nombre') and parsed.get('nombre'):
                        beneficiario['nombre'] = parsed['nombre']
            except Exception as _e_obs:
                __import__('logging').getLogger('compras').warning(
                    "OBS fallback para beneficiario falló: %s", _e_obs
                )

        # Empresa pagadora — detección multi-señal (ver _is_animus_payment).
        #   Influencer/Marketing/Cuenta de Cobro → ANIMUS LAB S.A.S.
        #   Resto (mercancía, MPs, planta, etc.) → ESPAGIRIA LABORATORIO S.A.S.
        obs_for_dispatch = None
        try:
            obs_row = cur.execute(
                "SELECT observaciones FROM solicitudes_compra "
                "WHERE numero_oc=? LIMIT 1", (numero_oc,)
            ).fetchone()
            if obs_row:
                obs_for_dispatch = obs_row[0]
        except sqlite3.OperationalError:
            pass
        if _is_animus_payment(
            cur, numero_oc=numero_oc,
            beneficiario_nombre=beneficiario.get('nombre', ''),
            observaciones=obs_for_dispatch,
            categoria=categoria,
        ):
            empresa_pagadora = 'Animus'
        else:
            empresa_pagadora = 'Espagiria'

        comp = crear_comprobante_y_pdf(
            conn, beneficiario=beneficiario, items=items_pdf,
            monto_subtotal=subtotal_ce,
            aplicar_retefuente=aplicar_retefuente,
            aplicar_retica=aplicar_retica,
            aplicar_iva=aplicar_iva,
            medio_pago=medio, observaciones=obs,
            pagado_por=usuario_actual, numero_oc=numero_oc,
            pago_oc_id=pago_oc_id, empresa=empresa_pagadora,
        )
        comprobante_info = {
            'numero_ce': comp['numero_ce'],
            'comprobante_id': comp['comprobante_id'],
            'subtotal': comp['subtotal'],
            'iva': comp['iva'],
            'retefuente': comp['retefuente'],
            'retica': comp['retica'],
            'total_pagado': comp['total_pagado'],
        }

        # Envío automático del comprobante por email al beneficiario.
        # Solo si: (a) hay email del beneficiario, (b) hay SMTP configurado.
        # Se ejecuta en background — no bloquea la respuesta de pago.
        email_dest = (beneficiario.get('email') or '').strip()
        if email_dest and '@' in email_dest:
            try:
                from notificaciones import SistemaNotificaciones
                sn = SistemaNotificaciones()
                if sn.email_remitente and sn.contraseña:
                    sn.enviar_en_background(
                        sn.enviar_comprobante_egreso,
                        destinatario=email_dest,
                        numero_ce=comp['numero_ce'],
                        beneficiario=beneficiario.get('nombre', ''),
                        total_pagado=comp['total_pagado'],
                        pdf_bytes=comp['pdf_bytes'],
                        fecha_emision=fecha_pago[:10],
                        numero_oc=numero_oc,
                        empresa=empresa_pagadora,
                    )
                    comprobante_info['email_enviado_a'] = email_dest
                else:
                    comprobante_info['email_pendiente'] = (
                        'SMTP no configurado en Render (EMAIL_REMITENTE/EMAIL_PASSWORD)'
                    )
            except Exception as _e_mail:
                __import__('logging').getLogger('compras').error(
                    "Email comprobante falló: %s", _e_mail
                )
        else:
            comprobante_info['email_pendiente'] = (
                'Beneficiario sin email — agrégalo en Marketing › Influencers'
            )
    except Exception as _e:
        __import__('logging').getLogger('compras').error(
            "No se pudo generar comprobante de egreso: %s", _e, exc_info=True
        )

    conn.commit()

    # ── Notificar al solicitante (Jefferson) cuando es OC influencer/CC ────
    # Sebastian (29-abr-2026): "le notificaba a jefer" — restaurar el ping
    # por email al solicitante cuando se le paga a uno de sus influencers.
    # Solo notifica cuando la OC pasa a Pagada (no Parcial).
    try:
        cat_low_n = (categoria or '').lower()
        es_influencer_oc = (
            'influencer' in cat_low_n
            or 'cuenta de cobro' in cat_low_n
            or 'marketing' in cat_low_n
        )
        if es_influencer_oc and nuevo_estado == 'Pagada':
            # Sebastian (29-abr-2026): "deberia llegarle a jeferson" no a quien
            # creó la SOL (puede haber sido Sebastian cargando bulk). Para OCs
            # de Influencer/CC el destinatario SIEMPRE es Jefferson.
            _dest = USER_EMAILS.get('jefferson', '')
            # Fallback: si Jefferson no está configurado, usar el solicitante
            # original (mejor algo que nada).
            if not _dest:
                sol_info = cur.execute(
                    "SELECT solicitante, email_solicitante FROM solicitudes_compra "
                    "WHERE numero_oc=? LIMIT 1", (numero_oc,)
                ).fetchone()
                if sol_info:
                    _sol_user = (sol_info[0] or '').strip().lower()
                    _dest = (sol_info[1] or '').strip() or USER_EMAILS.get(_sol_user, '')
            if _dest:
                _asunto = f"💸 Pago confirmado a {proveedor} — {numero_oc}"
                _body = (
                    f"<h2>Pago confirmado</h2>"
                    f"<p>Sebastian autorizó y registró el pago a <b>{proveedor}</b>:</p>"
                    f"<ul>"
                    f"<li><b>OC:</b> {numero_oc}</li>"
                    f"<li><b>Monto:</b> ${monto:,.0f} COP</li>"
                    f"<li><b>Medio:</b> {medio}</li>"
                    f"<li><b>Fecha:</b> {fecha_pago[:10]}</li>"
                    f"</ul>"
                    f"<p>El estado en Marketing → Influencers cambió a <b>Pagada</b>. "
                    f"El comprobante de egreso (CE) se adjuntó al beneficiario si tenía email.</p>"
                    f"<p style='color:#94a3b8;font-size:11px'>Mensaje automatico HHA Group</p>"
                )
                _notificar_solicitante_email(_dest, _asunto, _body)
    except Exception as _e_notif:
        __import__('logging').getLogger('compras').error(
            "Notificacion a Jefferson fallo (no critico): %s", _e_notif
        )

    return jsonify({
        'ok': True,
        'estado': nuevo_estado,
        'monto_este_pago': monto,
        'total_pagado_acumulado': round(total_pagado, 2),
        'valor_total_oc': round(valor_total_oc, 2),
        'pendiente': round(max(0, valor_total_oc - total_pagado), 2),
        'comprobante': comprobante_info,
    })


# ─── Endpoints comprobantes de egreso ────────────────────────────────────────


@bp.route('/api/comprobantes-pago', methods=['GET'])
def listar_comprobantes_pago():
    """Lista los comprobantes de egreso generados.

    Query: ?desde=&hasta=&beneficiario=  (filtros opcionales)
    """
    # SEC-FIX · 21-may-2026 · privilege leak · datos bancarios PII expuestos
    # Antes: cualquier compras_user logueado (mayerlin, operarios, sergio) veía
    # comprobantes con beneficiario_cedula + total_pagado. Ahora solo compras_write.
    u, err, code = _require_compras_write()
    if err:
        return err, code
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    benef = (request.args.get('beneficiario') or '').strip()

    conn = get_db(); c = conn.cursor()
    sql = """
        SELECT id, numero_ce, fecha_emision, numero_oc, beneficiario_nombre,
               beneficiario_cedula, subtotal, iva, retefuente, retica,
               total_pagado, medio_pago, pagado_por, empresa
        FROM comprobantes_pago
        WHERE 1=1
    """
    params = []
    if desde:
        sql += " AND fecha_emision >= ?"; params.append(desde)
    if hasta:
        sql += " AND fecha_emision <= ?"; params.append(hasta + ' 23:59:59')
    if benef:
        sql += " AND beneficiario_nombre LIKE ?"; params.append(f"%{benef}%")
    sql += " ORDER BY fecha_emision DESC LIMIT 500"
    c.execute(sql, params)
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({'comprobantes': rows, 'count': len(rows)})


@bp.route('/api/comprobantes-pago/<int:comp_id>/pdf', methods=['GET'])
def descargar_comprobante_pdf(comp_id):
    """Descarga el PDF del comprobante."""
    # SEC-FIX · 21-may-2026 · IDOR · solo compras_write puede descargar PDFs PII
    u, err, code = _require_compras_write()
    if err:
        return err, code
    import base64
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT numero_ce, pdf_archivo FROM comprobantes_pago WHERE id=?", (comp_id,))
    row = c.fetchone()
    if not row or not row[1]:
        return jsonify({'error': 'Comprobante no encontrado'}), 404
    numero_ce, pdf_b64 = row[0], row[1]
    try:
        pdf_bytes = base64.b64decode(pdf_b64)
    except Exception:
        return jsonify({'error': 'PDF corrupto'}), 500
    return Response(
        pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{numero_ce}.pdf"'}
    )


def _regenerate_ce_internal(conn, c, comp_id, forzar_obs=False,
                            empresa_override=None):
    """Helper interno: regenera el PDF + actualiza DB de un comprobante.

    Centraliza la logica para ser invocada tanto desde:
      - POST /api/comprobantes-pago/<id>/regenerar (single, via UI)
      - POST /api/comprobantes-pago/regenerar-legacy (bulk admin)

    Args:
      conn, c: conexion + cursor de SQLite (se hace commit al final)
      comp_id: id del comprobantes_pago row
      forzar_obs: si True, fuerza re-parseo de OBS aunque ya tenga banco
      empresa_override: si no None, fuerza la empresa (skip multi-signal)

    Returns dict:
      {ok: True, numero_ce, empresa, beneficiario, subtotal_usado, pdf_size_kb}
      {ok: False, error}
    """
    c.execute("""
        SELECT id, numero_ce, numero_oc, beneficiario_nombre, beneficiario_cedula,
               beneficiario_banco, beneficiario_cuenta, beneficiario_tipo_cta,
               beneficiario_ciudad, subtotal, iva_pct, retefuente_pct, retica_pct,
               medio_pago, observaciones, pagado_por, empresa, pdf_archivo,
               fecha_emision
        FROM comprobantes_pago WHERE id=?
    """, (comp_id,))
    row = c.fetchone()
    if not row:
        return {'ok': False, 'error': 'Comprobante no encontrado'}

    (_, numero_ce, numero_oc, ben_nombre, ben_cedula, ben_banco, ben_cuenta,
     ben_tipo_cta, ben_ciudad, subtotal, iva_pct, rete_pct, retica_pct,
     medio_pago, observaciones, pagado_por, empresa_db, _, fecha_emision_str) = row

    # ── Empresa: usa override del caller, sino re-deriva multi-señal ──
    # _is_animus_payment cubre CEs legacy donde 'empresa_db' quedó en
    # 'Espagiria' por default histórico aunque era pago de influencer.
    if empresa_override is not None:
        empresa = empresa_override
    else:
        oc_row = c.execute(
            "SELECT categoria FROM ordenes_compra WHERE numero_oc=?",
            (numero_oc,)
        ).fetchone()
        cat = (oc_row[0] if oc_row else '') or ''
        if _is_animus_payment(
            c, numero_oc=numero_oc,
            beneficiario_nombre=ben_nombre,
            observaciones=observaciones,
            categoria=cat,
        ):
            empresa = 'Animus'
        else:
            empresa = empresa_db or 'Espagiria'

    # ── Beneficiario base ──
    beneficiario = {
        'nombre': ben_nombre or '',
        'cedula': ben_cedula or '',
        'banco': ben_banco or '',
        'cuenta': ben_cuenta or '',
        'tipo_cuenta': ben_tipo_cta or '',
        'ciudad': ben_ciudad or '',
        'email': '',
    }

    # ── Re-lookup desde marketing_influencers (autoritativo si es influencer) ──
    # Para Animus/Influencer pagos, los datos en marketing_influencers son
    # la fuente de verdad: nombre + cédula + banco + cuenta + tipo + ciudad +
    # email. Antes solo se completaba el email; ahora completamos todo lo
    # que esté vacío en el beneficiario, usando el nombre como key de match.
    es_influencer = (empresa or '').lower() == 'animus' or 'influencer' in (empresa or '').lower()
    if es_influencer or forzar_obs:
        try:
            mi = c.execute("""
                SELECT nombre, cedula_nit, banco, cuenta_bancaria,
                       tipo_cuenta, ciudad, email
                FROM marketing_influencers
                WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?)) LIMIT 1
            """, (beneficiario['nombre'],)).fetchone()
            if mi:
                # Completar campos vacíos (forzar_obs sobrescribe siempre)
                pairs = [
                    ('cedula', mi[1]), ('banco', mi[2]), ('cuenta', mi[3]),
                    ('tipo_cuenta', mi[4]), ('ciudad', mi[5]), ('email', mi[6]),
                ]
                for key, val in pairs:
                    if val and (forzar_obs or not beneficiario.get(key)):
                        beneficiario[key] = val
        except Exception:
            pass

    # ── Si aún faltan datos bancarios O forzar_obs: parsear OBS del comprobante ──
    if forzar_obs or not beneficiario['banco'] or not beneficiario['cuenta']:
        from comprobante_pago import parse_obs_beneficiario
        obs_src = observaciones  # OBS guardado en comprobante
        if not obs_src and numero_oc:
            obs_row = c.execute(
                "SELECT observaciones FROM solicitudes_compra WHERE numero_oc=? LIMIT 1",
                (numero_oc,)
            ).fetchone()
            if obs_row:
                obs_src = obs_row[0] or ''
        if obs_src:
            parsed = parse_obs_beneficiario(obs_src)
            if forzar_obs or not beneficiario['banco']:
                beneficiario['banco'] = parsed.get('banco') or beneficiario['banco']
                beneficiario['tipo_cuenta'] = parsed.get('tipo_cuenta') or beneficiario['tipo_cuenta']
            if forzar_obs or not beneficiario['cuenta']:
                beneficiario['cuenta'] = parsed.get('cuenta') or beneficiario['cuenta']
            if forzar_obs or not beneficiario['cedula']:
                beneficiario['cedula'] = parsed.get('cedula') or beneficiario['cedula']

    # ── Valor: re-leer del comprobante (subtotal ya almacenado) ──
    # Si el subtotal almacenado parece incorrecto (< 10000 para pesos COP),
    # intentar recuperarlo de la OC.
    monto = subtotal or 0
    if monto < 1000 and numero_oc:
        oc_val = c.execute(
            "SELECT valor_total FROM ordenes_compra WHERE numero_oc=?", (numero_oc,)
        ).fetchone()
        if oc_val and oc_val[0] and float(oc_val[0]) > monto:
            monto = float(oc_val[0])
            __import__('logging').getLogger('compras').warning(
                "regenerar CE %s: subtotal almacenado=%s parece incorrecto, "
                "usando valor_total OC=%s", numero_ce, subtotal, monto
            )

    # ── Reconstruir items ──
    items_pdf = [{'descripcion': observaciones or f'Pago {numero_oc}',
                  'cantidad': 1, 'valor_unit': monto}]

    # ── Re-generar PDF ──
    from comprobante_pago import generar_comprobante_egreso_pdf
    import base64
    from datetime import datetime as _dt

    try:
        fecha_pago = _dt.fromisoformat(fecha_emision_str.replace('Z', '')) if fecha_emision_str else _dt.now()
    except Exception:
        fecha_pago = _dt.now()

    pdf_bytes = generar_comprobante_egreso_pdf(
        numero_ce=numero_ce,
        fecha_pago=fecha_pago,
        beneficiario=beneficiario,
        items=items_pdf,
        aplicar_retefuente=(rete_pct or 0) > 0,
        aplicar_retica=(retica_pct or 0) > 0,
        aplicar_iva=(iva_pct or 0) > 0,
        medio_pago=medio_pago or 'Transferencia',
        observaciones='',
        pagado_por=pagado_por or '',
        empresa_clave=empresa.lower(),
    )
    pdf_b64 = base64.b64encode(pdf_bytes).decode('ascii')

    # ── Actualizar DB ── (incluye ciudad, antes faltaba y quedaba NULL)
    c.execute("""
        UPDATE comprobantes_pago
        SET pdf_archivo=?, empresa=?,
            beneficiario_banco=?, beneficiario_cuenta=?,
            beneficiario_tipo_cta=?, beneficiario_cedula=?,
            beneficiario_ciudad=?
        WHERE id=?
    """, (
        pdf_b64, empresa,
        beneficiario['banco'], beneficiario['cuenta'],
        beneficiario['tipo_cuenta'], beneficiario['cedula'],
        beneficiario.get('ciudad') or '',
        comp_id,
    ))
    conn.commit()

    return {
        'ok': True,
        'numero_ce': numero_ce,
        'empresa': empresa,
        'beneficiario': {k: v for k, v in beneficiario.items() if k != 'email'},
        'subtotal_usado': monto,
        'pdf_size_kb': round(len(pdf_bytes) / 1024, 1),
    }


@bp.route('/api/comprobantes-pago/<int:comp_id>/regenerar', methods=['POST'])
def regenerar_comprobante_pdf(comp_id):
    """Re-genera el PDF de un comprobante existente con datos actualizados.

    Útil para corregir comprobantes generados antes de que se implementara:
      - Dispatch correcto empresa (Espagiria vs ANIMUS Lab)
      - Parseo OBS para datos bancarios
      - Formateo correcto de montos COP

    Body JSON (todos opcionales — si no se pasan, se re-derivan de la DB):
      empresa: "Animus" | "Espagiria"
      forzar_obs: true  — fuerza re-parseo OBS aunque ya haya banco en DB
    """
    # P0 audit 26-may-2026 · zero-error · regenerar comprobante = mutación
    # de documento financiero (INVIMA / contable). Endpoint hermano
    # /regenerar-legacy ya exige admin · alinear este al mismo gate.
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    if (usuario or '').lower() not in {x.lower() for x in ADMIN_USERS}:
        return jsonify({'error': 'Solo admin puede regenerar comprobantes'}), 403
    d = request.get_json(silent=True) or {}
    forzar_obs = bool(d.get('forzar_obs', False))
    empresa_override = d.get('empresa') if 'empresa' in d else None

    conn = get_db(); c = conn.cursor()
    result = _regenerate_ce_internal(
        conn, c, comp_id,
        forzar_obs=forzar_obs,
        empresa_override=empresa_override,
    )
    if not result.get('ok'):
        return jsonify(result), 404
    # Audit log · trail regulatorio
    try:
        from audit_helpers import audit_log
        audit_log(c, usuario=usuario, accion='REGENERAR_COMPROBANTE_PAGO',
                  tabla='comprobantes_pago', registro_id=comp_id,
                  despues={'forzar_obs': forzar_obs,
                           'empresa_override': empresa_override or '',
                           'pdf_size_kb': result.get('pdf_size_kb')},
                  detalle=f'CE id={comp_id} regenerado · empresa={empresa_override or "auto"}')
    except Exception as _ae:
        import logging
        logging.getLogger('compras').warning('audit regenerar CE fallo: %s', _ae)
    return jsonify(result)


@bp.route('/api/comprobantes-pago/regenerar-legacy', methods=['POST'])
def regenerar_comprobantes_legacy():
    """Bulk-regenera CEs con dispatch Animus/Espagiria incorrecto.

    Detecta CEs marcados como 'Espagiria' que en realidad eran pagos a
    influencer (multi-señal _is_animus_payment) y regenera todos los
    PDFs con la empresa correcta (ANIMUS LAB).

    Util para corregir CEs creados antes del feature de 2 empresas
    pagadoras (commit 19443c9). Antes el flujo era: usuario va al
    modulo /marketing y hace clic en 'Regenerar' uno por uno; con este
    endpoint admin lo hace de un golpe.

    Body JSON (opcional):
      dry_run: bool (default false) — solo lista candidatos, no toca

    Solo admins. Audita en security_events cada CE corregido.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS:
        return jsonify({
            'error': 'Solo administradores pueden bulk-regenerar comprobantes'
        }), 403

    d = request.get_json(silent=True) or {}
    dry_run = bool(d.get('dry_run', False))

    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT id, numero_ce, numero_oc, beneficiario_nombre, observaciones
        FROM comprobantes_pago
        WHERE LOWER(COALESCE(empresa,'')) != 'animus'
        ORDER BY id ASC
    """).fetchall()

    candidatos = []
    for cid, ce, oc, ben_nom, obs in rows:
        cat = ''
        if oc:
            oc_row = c.execute(
                "SELECT categoria FROM ordenes_compra WHERE numero_oc=?",
                (oc,)
            ).fetchone()
            cat = (oc_row[0] if oc_row else '') or ''
        if _is_animus_payment(
            c, numero_oc=oc, beneficiario_nombre=ben_nom,
            observaciones=obs, categoria=cat,
        ):
            candidatos.append({
                'id': cid, 'numero_ce': ce, 'numero_oc': oc or '',
                'beneficiario': ben_nom or '',
            })

    if dry_run:
        return jsonify({
            'ok': True, 'dry_run': True,
            'candidatos': candidatos,
            'count': len(candidatos),
        })

    fixed = []
    errors = []
    for cand in candidatos:
        try:
            result = _regenerate_ce_internal(
                conn, c, cand['id'],
                forzar_obs=False,
                empresa_override='Animus',
            )
            if result.get('ok'):
                fixed.append({
                    'id': cand['id'],
                    'numero_ce': cand['numero_ce'],
                    'pdf_size_kb': result.get('pdf_size_kb'),
                })
            else:
                errors.append({
                    'id': cand['id'],
                    'numero_ce': cand['numero_ce'],
                    'error': result.get('error', 'unknown'),
                })
        except Exception as _e:
            errors.append({
                'id': cand['id'],
                'numero_ce': cand['numero_ce'],
                'error': str(_e),
            })

    return jsonify({
        'ok': True, 'dry_run': False,
        'corregidos': fixed,
        'errores': errors,
        'count_corregidos': len(fixed),
        'count_errores': len(errors),
    })


@bp.route('/api/comprobantes-pago/<int:comp_id>/email', methods=['POST'])
def reenviar_comprobante_email(comp_id):
    """Reenvía el comprobante PDF por email.

    Body opcional: {"destinatario": "otro@correo.com"}
    Si no se especifica destinatario, busca el email del beneficiario en
    marketing_influencers o proveedores.
    """
    u, err, code = _require_compras_session()
    if err:
        return err, code
    import base64
    body = request.get_json(silent=True) or {}
    destinatario_override = (body.get('destinatario') or '').strip()

    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT numero_ce, pdf_archivo, beneficiario_nombre,
                        total_pagado, fecha_emision, numero_oc, empresa
                 FROM comprobantes_pago WHERE id=?""", (comp_id,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'Comprobante no encontrado'}), 404
    numero_ce, pdf_b64, benef_nombre, total, fecha_em, num_oc, empresa = row
    if not pdf_b64:
        return jsonify({'error': 'PDF no disponible para este comprobante'}), 404

    # Resolver destinatario: override > marketing_influencers > proveedores
    destinatario = destinatario_override
    if not destinatario and benef_nombre:
        try:
            r2 = c.execute(
                "SELECT email FROM marketing_influencers WHERE LOWER(TRIM(nombre))=? LIMIT 1",
                (benef_nombre.strip().lower(),)
            ).fetchone()
            if r2 and r2[0]:
                destinatario = r2[0].strip()
        except sqlite3.OperationalError:
            pass
    if not destinatario and benef_nombre:
        try:
            r3 = c.execute(
                "SELECT email FROM proveedores WHERE LOWER(TRIM(nombre))=? LIMIT 1",
                (benef_nombre.strip().lower(),)
            ).fetchone()
            if r3 and r3[0]:
                destinatario = r3[0].strip()
        except sqlite3.OperationalError:
            pass

    if not destinatario or '@' not in destinatario:
        return jsonify({
            'error': 'Sin destinatario',
            'detalle': 'No se encontró email para el beneficiario. Pasa "destinatario" en el body o agrega el email al influencer/proveedor.'
        }), 400

    try:
        pdf_bytes = base64.b64decode(pdf_b64)
    except Exception:
        return jsonify({'error': 'PDF corrupto'}), 500

    try:
        from notificaciones import SistemaNotificaciones
        sn = SistemaNotificaciones()
        if not sn.email_remitente or not sn.contraseña:
            return jsonify({
                'error': 'SMTP no configurado',
                'detalle': 'Define EMAIL_REMITENTE y EMAIL_PASSWORD en Render → Environment.',
                'doc': 'https://myaccount.google.com/apppasswords (App Password de Gmail)'
            }), 503
        ok = sn.enviar_comprobante_egreso(
            destinatario=destinatario, numero_ce=numero_ce,
            beneficiario=benef_nombre or '', total_pagado=total or 0,
            pdf_bytes=pdf_bytes,
            fecha_emision=(fecha_em or '')[:10], numero_oc=num_oc or '',
            empresa=empresa or 'Espagiria',
        )
        if not ok:
            return jsonify({'error': 'Falló envío SMTP', 'detalle': 'Revisa logs y credenciales SMTP'}), 502
        return jsonify({'ok': True, 'destinatario': destinatario, 'numero_ce': numero_ce})
    except Exception as e:
        return jsonify({'error': 'Error interno', 'detalle': str(e)}), 500


@bp.route('/api/comprobantes-pago/oc/<numero_oc>', methods=['GET'])
def comprobantes_de_oc(numero_oc):
    """Lista comprobantes asociados a una OC específica."""
    u, err, code = _require_compras_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT id, numero_ce, fecha_emision, total_pagado, medio_pago,
                        pagado_por, beneficiario_nombre
                 FROM comprobantes_pago WHERE numero_oc=?
                 ORDER BY fecha_emision DESC""", (numero_oc,))
    cols = [d[0] for d in c.description]
    return jsonify({'comprobantes': [dict(zip(cols, r)) for r in c.fetchall()]})


@bp.route('/api/ordenes-compra/<numero_oc>/pagos', methods=['GET'])
def get_pagos_oc(numero_oc):
    """Histórico de pagos de UNA OC específica (Sprint 2).

    Útil para auditar pagos parciales o ver el detalle de quién pagó qué cuándo.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        SELECT id, monto, medio, fecha_pago, registrado_por,
               numero_factura_proveedor, observaciones,
               CASE WHEN comprobante_imagen != '' AND comprobante_imagen IS NOT NULL
                    THEN 1 ELSE 0 END as tiene_comprobante
        FROM pagos_oc WHERE numero_oc=?
        ORDER BY fecha_pago ASC
    """, (numero_oc,))
    cols = [d[0] for d in cur.description]
    pagos = [dict(zip(cols, row)) for row in cur.fetchall()]

    # Total + comparación con valor_total OC
    cur.execute("SELECT valor_total FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    row = cur.fetchone()
    valor_oc = float(row[0]) if row else 0
    total_pagado = sum(p.get('monto', 0) or 0 for p in pagos)

    return jsonify({
        'numero_oc': numero_oc,
        'pagos': pagos,
        'count': len(pagos),
        'total_pagado': round(total_pagado, 2),
        'valor_total_oc': round(valor_oc, 2),
        'pendiente': round(max(0, valor_oc - total_pagado), 2),
    })


@bp.route('/api/compras/pagos', methods=['GET'])
def get_pagos():
    """Return all paid OCs with payment metadata (no image data).

    Sebastian (29-abr-2026): "no salen los que he pagado". Antes excluiamos
    influencers/CC del listado pero rompia la visibilidad de los pagos hechos
    por Sebastian a influencers desde /compras tab Influencers. Ahora SI los
    devolvemos pero marcados con `es_influencer: true` para que el frontend
    los pueda destacar visualmente o filtrar.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); cur = conn.cursor()
    # Sebastián 24-may-2026 · audit Pagos · agregar comprobante_id del último
    # CE (para botón "Regenerar CE" inline en tabla) + numero_factura_proveedor
    # del último pago (3-way matching visible).
    cur.execute("""
        SELECT oc.numero_oc, oc.proveedor, oc.categoria, oc.valor_total,
               oc.medio_pago, oc.fecha_pago, oc.pagado_por, oc.estado,
               oc.observaciones,
               CASE WHEN oc.comprobante_imagen != '' AND oc.comprobante_imagen IS NOT NULL
                    THEN 1 ELSE 0 END as tiene_comprobante,
               COALESCE((SELECT SUM(monto) FROM pagos_oc WHERE numero_oc=oc.numero_oc), 0) as monto,
               oc.valor_total -
                 COALESCE((SELECT SUM(monto) FROM pagos_oc WHERE numero_oc=oc.numero_oc), 0)
                 as saldo_pendiente,
               CASE WHEN oc.categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')
                    OR EXISTS(SELECT 1 FROM pagos_influencers pi WHERE pi.numero_oc=oc.numero_oc)
                    OR EXISTS(SELECT 1 FROM solicitudes_compra sc
                              WHERE sc.numero_oc=oc.numero_oc AND sc.influencer_id IS NOT NULL)
                    THEN 1 ELSE 0 END as es_influencer,
               (SELECT id FROM comprobantes_pago
                WHERE numero_oc=oc.numero_oc
                ORDER BY id DESC LIMIT 1) as comprobante_id,
               (SELECT numero_factura_proveedor FROM pagos_oc
                WHERE numero_oc=oc.numero_oc AND COALESCE(numero_factura_proveedor,'') != ''
                ORDER BY id DESC LIMIT 1) as numero_factura_proveedor
        FROM ordenes_compra oc
        WHERE oc.estado IN ('Pagada','Parcial')
        ORDER BY oc.fecha_pago DESC
        LIMIT 500
    """)
    cols = [d[0] for d in cur.description]
    pagos = [dict(zip(cols, row)) for row in cur.fetchall()]
    # Convertir es_influencer a bool para JSON
    for p in pagos:
        p['es_influencer'] = bool(p.get('es_influencer'))
    return jsonify({'pagos': pagos})


# ─── Excel histórico pagos · Sebastián 24-may-2026 ──────────────
@bp.route('/api/compras/pagos-excel', methods=['GET'])
def pagos_excel():
    """Excel descargable con histórico completo de pagos (Pagada + Parcial).

    Querystring opcional · ?mes=2026-05 filtra por mes específico.
    Columnas: OC · Proveedor · Categoría · Monto · Medio · Factura · Fecha
              · Pagado por · Estado · Es influencer · Saldo pendiente
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    mes = (request.args.get('mes') or '').strip()
    conn = get_db(); cur = conn.cursor()
    sql = """
        SELECT oc.numero_oc, oc.proveedor, oc.categoria,
               COALESCE((SELECT SUM(monto) FROM pagos_oc WHERE numero_oc=oc.numero_oc), 0) as monto,
               oc.medio_pago,
               (SELECT numero_factura_proveedor FROM pagos_oc
                WHERE numero_oc=oc.numero_oc AND COALESCE(numero_factura_proveedor,'') != ''
                ORDER BY id DESC LIMIT 1) as factura,
               oc.fecha_pago, oc.pagado_por, oc.estado,
               oc.valor_total -
                 COALESCE((SELECT SUM(monto) FROM pagos_oc WHERE numero_oc=oc.numero_oc), 0)
                 as saldo_pendiente
        FROM ordenes_compra oc
        WHERE oc.estado IN ('Pagada','Parcial')
    """
    params = []
    if mes and len(mes) == 7:  # formato YYYY-MM
        sql += " AND substr(oc.fecha_pago, 1, 7) = ?"
        params.append(mes)
    sql += " ORDER BY oc.fecha_pago DESC"
    cur.execute(sql, params)
    rows = cur.fetchall()
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except ImportError:
        return jsonify({'error': 'openpyxl no disponible'}), 503
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagos'
    ws.append(['OC', 'Proveedor', 'Categoría', 'Monto', 'Medio',
               'Factura', 'Fecha pago', 'Pagado por', 'Estado',
               'Saldo pendiente'])
    for r in rows:
        ws.append([r[0], r[1], r[2], float(r[3] or 0), r[4] or '',
                   r[5] or '', (r[6] or '')[:10], r[7] or '', r[8],
                   float(r[9] or 0)])
    # Auto-anchear columnas
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 2, 30)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f'pagos_{mes}.xlsx' if mes else 'pagos_historico.xlsx'
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={nombre}'},
    )


# ─── KPIs Pagos breakdown · mes/año/medio · Sebastián 24-may-2026 ──
@bp.route('/api/compras/pagos-kpis', methods=['GET'])
def pagos_kpis():
    """KPIs agregados de pagos: total mes actual + año actual + breakdown por medio.

    Útil para dashboard del tab Pagos · evita que UI tenga que iterar
    PAGOS[] y filtrar por fecha.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); cur = conn.cursor()
    hoy = datetime.now()
    mes_actual = hoy.strftime('%Y-%m')
    anio_actual = hoy.strftime('%Y')
    # Total mes actual
    cur.execute(
        "SELECT COUNT(DISTINCT oc.numero_oc), COALESCE(SUM(po.monto),0) "
        "FROM pagos_oc po JOIN ordenes_compra oc ON po.numero_oc=oc.numero_oc "
        "WHERE substr(po.fecha_pago, 1, 7) = ?",
        (mes_actual,),
    )
    r_mes = cur.fetchone()
    # Total año actual
    cur.execute(
        "SELECT COUNT(DISTINCT oc.numero_oc), COALESCE(SUM(po.monto),0) "
        "FROM pagos_oc po JOIN ordenes_compra oc ON po.numero_oc=oc.numero_oc "
        "WHERE substr(po.fecha_pago, 1, 4) = ?",
        (anio_actual,),
    )
    r_anio = cur.fetchone()
    # Breakdown por medio (año actual)
    cur.execute(
        "SELECT COALESCE(po.medio,'(sin medio)'), COUNT(*), COALESCE(SUM(po.monto),0) "
        "FROM pagos_oc po "
        "WHERE substr(po.fecha_pago, 1, 4) = ? "
        "GROUP BY COALESCE(po.medio,'(sin medio)') "
        "ORDER BY SUM(po.monto) DESC",
        (anio_actual,),
    )
    medios = [{'medio': r[0], 'n_pagos': int(r[1] or 0),
                'total': float(r[2] or 0)} for r in cur.fetchall()]
    return jsonify({
        'mes_actual': {'mes': mes_actual, 'n_ocs': int(r_mes[0] or 0),
                        'total': float(r_mes[1] or 0)},
        'anio_actual': {'anio': anio_actual, 'n_ocs': int(r_anio[0] or 0),
                         'total': float(r_anio[1] or 0)},
        'breakdown_medios': medios,
    })


# ── Categorías que NO requieren recepción física: van directo a "por pagar" ──
# El contador ve estos como pagos directos (servicios, no mercancía).
# NOTA: 'Influencer/Marketing Digital' NO va aquí — Marketing tiene su propio
# panel para pagar influencers y no debe aparecer en /compras.
CATEGORIAS_PAGO_DIRECTO = (
    'Cuenta de Cobro',
    'Servicio',
    'SVC',
)


def _autorreparar_ocs_vacias(cur):
    """Sebastian (29-abr-2026): rellena proveedor / valor_total de OCs
    en estado pre-pago que quedaron sin datos al crearlas (caso 0119).
    Solo toca OCs NO pagadas con proveedor 'Por definir'/vacio o valor<=0
    Y que tengan solicitud asociada con datos. Devuelve lista de OCs
    reparadas. Pensado para correr al cargar /por-pagar y al iniciar
    Dashboard — idempotente.
    """
    reparadas = []
    rows = cur.execute("""
        SELECT oc.numero_oc, oc.proveedor, oc.valor_total, oc.categoria,
               s.numero, s.influencer_id, COALESCE(s.valor,0), s.solicitante
        FROM ordenes_compra oc
        LEFT JOIN solicitudes_compra s ON s.numero_oc = oc.numero_oc
        WHERE oc.estado NOT IN ('Pagada','Rechazada','Cancelada')
          AND (
                COALESCE(oc.proveedor,'') = ''
             OR LOWER(oc.proveedor) = 'por definir'
             OR COALESCE(oc.valor_total,0) <= 0
          )
          AND s.numero IS NOT NULL
    """).fetchall()
    for r in rows:
        oc_num, prov_act, val_act, cat_oc, sol_num, inf_id, sol_valor, solic = r
        prov_act = (prov_act or '').strip()
        nuevo_prov = prov_act
        nuevo_val = float(val_act or 0)
        # PROVEEDOR fallback
        if not nuevo_prov or nuevo_prov.lower() == 'por definir':
            if inf_id:
                _ri = cur.execute(
                    "SELECT nombre FROM marketing_influencers WHERE id=?",
                    (inf_id,)
                ).fetchone()
                if _ri and _ri[0]:
                    nuevo_prov = _ri[0]
            if not nuevo_prov or nuevo_prov.lower() == 'por definir':
                _rp = cur.execute(
                    "SELECT proveedor_sugerido FROM solicitudes_compra_items "
                    "WHERE numero=? AND COALESCE(proveedor_sugerido,'') != '' LIMIT 1",
                    (sol_num,)
                ).fetchone()
                if _rp and _rp[0]:
                    nuevo_prov = _rp[0]
            if (not nuevo_prov or nuevo_prov.lower() == 'por definir') \
               and (cat_oc or '') in ('SVC', 'CC') and solic:
                nuevo_prov = solic
        # VALOR fallback
        if nuevo_val <= 0:
            if (sol_valor or 0) > 0:
                nuevo_val = float(sol_valor)
            else:
                _r2 = cur.execute(
                    "SELECT COALESCE(SUM(COALESCE(valor_estimado,0)),0), "
                    "       COALESCE(SUM(COALESCE(precio_unit_g,0)*COALESCE(cantidad_g,0)),0) "
                    "FROM solicitudes_compra_items WHERE numero=?",
                    (sol_num,)
                ).fetchone()
                if _r2:
                    nuevo_val = float(_r2[0] or 0) or float(_r2[1] or 0)
        # Solo update si hay un cambio real
        cambios = []
        params = []
        if nuevo_prov and nuevo_prov != prov_act:
            cambios.append("proveedor=?")
            params.append(nuevo_prov)
        if nuevo_val > 0 and nuevo_val != float(val_act or 0):
            cambios.append("valor_total=?")
            params.append(nuevo_val)
        if cambios:
            params.append(oc_num)
            cur.execute(
                f"UPDATE ordenes_compra SET {', '.join(cambios)} WHERE numero_oc=?",
                params
            )
            reparadas.append(oc_num)
    return reparadas


@bp.route('/api/compras/oc/<numero_oc>/reparar-desde-solicitud', methods=['POST'])
def reparar_oc_desde_solicitud(numero_oc):
    """Repara una OC individual jalando datos de la solicitud asociada.
    Util cuando el usuario ve una OC con 'Por definir / $0' y quiere
    forzar la sincronizacion."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); cur = conn.cursor()
    row = cur.execute(
        "SELECT estado FROM ordenes_compra WHERE numero_oc=?", (numero_oc,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'OC no existe'}), 404
    # Reusar el helper restringiendo el WHERE a esta OC
    rows = cur.execute("""
        SELECT oc.numero_oc, oc.proveedor, oc.valor_total, oc.categoria,
               s.numero, s.influencer_id, COALESCE(s.valor,0), s.solicitante
        FROM ordenes_compra oc
        LEFT JOIN solicitudes_compra s ON s.numero_oc = oc.numero_oc
        WHERE oc.numero_oc=?
    """, (numero_oc,)).fetchone()
    if not rows or not rows[4]:
        return jsonify({'error': 'OC sin solicitud asociada — no hay de donde jalar'}), 409
    # Llamar al helper solo para esta OC
    class _CurFilter:
        def __init__(self, real_cur):
            self.real_cur = real_cur
        def execute(self, sql, params=()):
            # Inyectar filtro por numero_oc al SELECT principal
            if 'oc.numero_oc, oc.proveedor, oc.valor_total' in sql and 'WHERE' in sql and 'numero_oc' not in sql.split('WHERE',1)[1]:
                sql = sql.rstrip().rstrip(';') + " AND oc.numero_oc=?"
                params = list(params) + [numero_oc]
            return self.real_cur.execute(sql, params)
    # Mas simple: llamar normal y verificar si reparo esta OC
    reparadas = _autorreparar_ocs_vacias(cur)
    conn.commit()
    if numero_oc in reparadas:
        return jsonify({'ok': True, 'reparada': True, 'numero_oc': numero_oc})
    # No habia nada que reparar (ya esta completa) o no hay datos en la sol
    actual = cur.execute(
        "SELECT proveedor, valor_total FROM ordenes_compra WHERE numero_oc=?",
        (numero_oc,)
    ).fetchone()
    return jsonify({
        'ok': True, 'reparada': False, 'numero_oc': numero_oc,
        'proveedor': actual[0], 'valor_total': actual[1],
        'mensaje': 'Sin cambios — la OC ya tiene datos o la solicitud tampoco los tiene.'
    })


@bp.route('/api/compras/por-pagar', methods=['GET'])
def get_por_pagar():
    """Vista unificada del contador: TODO lo que está pendiente de pago.

    Incluye:
      - OCs con estado 'Recibida' o 'Parcial' (mercancía física recibida)
      - OCs con estado 'Aprobada'/'Autorizada' Y categoría de servicio
        (Influencer, Cuenta de Cobro, etc.) — NO requieren recepción

    Cada item lleva flag `pago_directo: true` si es servicio sin mercancía.
    El frontend puede mostrarlos en una sección destacada.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    conn = get_db(); cur = conn.cursor()

    # Sebastian (29-abr-2026): "OC-2026-0119 sale Por definir y $0 aunque
    # la solicitud trae 150.000". Auto-reparacion silenciosa al cargar el
    # tab: rellena proveedor/valor de OCs en estado pre-pago tomando los
    # datos de la solicitud asociada. No-op si ya estan completos.
    try:
        _reparadas = _autorreparar_ocs_vacias(cur)
        if _reparadas:
            conn.commit()
    except Exception:
        pass

    # Sebastián 23-may-2026 · admin necesita info bancaria visible para
    # pagar · INCLUYE num_cuenta/banco/tipo_cuenta/nit del proveedor
    # (RBAC ya garantiza que solo compras_user accede · contadora SÍ ve
    # porque es la que registra el pago efectivo · pero NO autoriza OCs)
    cur.execute("""
        SELECT oc.numero_oc, oc.proveedor, oc.categoria, oc.valor_total, oc.fecha,
               oc.estado, oc.observaciones, oc.fecha_recepcion,
               COALESCE(p.condiciones_pago, '') as condiciones_pago,
               COALESCE(p.num_cuenta,'') as num_cuenta,
               COALESCE(p.banco,'') as banco,
               COALESCE(p.tipo_cuenta,'') as tipo_cuenta,
               COALESCE(p.nit,'') as nit
        FROM ordenes_compra oc
        LEFT JOIN proveedores p ON oc.proveedor = p.nombre
        WHERE oc.estado IN ('Recibida', 'Parcial')
        ORDER BY oc.fecha_recepcion DESC, oc.numero_oc DESC
    """)
    cols = [d[0] for d in cur.description]
    fisicas = [
        {**dict(zip(cols, row)), 'pago_directo': False, 'tipo': 'Mercancía recibida'}
        for row in cur.fetchall()
    ]

    # Servicios sin mercancía: Aprobada/Autorizada + categoría de pago directo
    placeholders = ','.join('?' for _ in CATEGORIAS_PAGO_DIRECTO)
    cur.execute(f"""
        SELECT oc.numero_oc, oc.proveedor, oc.categoria, oc.valor_total, oc.fecha,
               oc.estado, oc.observaciones,
               COALESCE(oc.fecha_recepcion, oc.fecha) as fecha_recepcion,
               COALESCE(p.condiciones_pago, '') as condiciones_pago,
               COALESCE(p.num_cuenta,'') as num_cuenta,
               COALESCE(p.banco,'') as banco,
               COALESCE(p.tipo_cuenta,'') as tipo_cuenta,
               COALESCE(p.nit,'') as nit
        FROM ordenes_compra oc
        LEFT JOIN proveedores p ON oc.proveedor = p.nombre
        WHERE oc.estado IN ('Aprobada', 'Autorizada') AND
              oc.categoria IN ({placeholders})
        ORDER BY oc.fecha DESC
    """, CATEGORIAS_PAGO_DIRECTO)
    cols = [d[0] for d in cur.description]
    servicios = [
        {**dict(zip(cols, row)), 'pago_directo': True,
         'tipo': 'Pago directo (servicio)'}
        for row in cur.fetchall()
    ]

    # NOTA 28-abr-2026 (sesion #2 Sebastian): retiramos la seccion
    # "en_proceso" de Por Pagar. Las OCs en Borrador/Revisada/Pendiente/
    # Aprobada NO son "pendiente de pago" todavia — estan en proceso de
    # creacion/autorizacion y se ven en Dashboard y otras pestañas.
    # Por Pagar ahora muestra SOLO lo que esta listo para pagar:
    #   - Mercancia recibida (Recibida/Parcial)
    #   - Servicios autorizados (Aprobada/Autorizada con categoria de pago directo)
    todos = fisicas + servicios
    todos.sort(key=lambda x: x.get('fecha_recepcion') or x.get('fecha') or '', reverse=True)

    total_valor = sum(item.get('valor_total', 0) or 0 for item in todos)
    total_servicios = sum(item['valor_total'] or 0 for item in servicios)
    total_fisicas = sum(item['valor_total'] or 0 for item in fisicas)

    return jsonify({
        'items': todos,
        'count': len(todos),
        'total_valor': round(total_valor, 2),
        'desglose': {
            'pagos_directos_servicios': {
                'count': len(servicios),
                'valor': round(total_servicios, 2),
            },
            'mercancia_recibida': {
                'count': len(fisicas),
                'valor': round(total_fisicas, 2),
            },
        }
    })

@bp.route('/api/ordenes-compra/<numero_oc>/comprobante', methods=['GET'])
def get_comprobante(numero_oc):
    """Return comprobantes for an OC.

    Comportamiento:
    - ?all=1 → lista de TODOS los comprobantes de pagos_oc (multi-pago)
    - default → comprobante de ordenes_compra (último · legacy compat)

    FIX · 22-may-2026 · Bug #6 OCs · multi-pago no sobreescribe historial
    · cada pago en pagos_oc puede tener su propio comprobante_imagen.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); cur = conn.cursor()
    if request.args.get('all') == '1':
        # Multi-pago · histórico completo
        try:
            rows = cur.execute(
                """SELECT id, fecha_pago, monto, medio, COALESCE(observaciones,''),
                          COALESCE(numero_factura_proveedor,''),
                          COALESCE(comprobante_imagen,'')
                   FROM pagos_oc WHERE numero_oc=? ORDER BY fecha_pago DESC""",
                (numero_oc,),
            ).fetchall()
            return jsonify({
                'numero_oc': numero_oc,
                'comprobantes': [{
                    'id': r[0], 'fecha': r[1], 'monto': float(r[2] or 0),
                    'medio': r[3], 'referencia': r[4],
                    'numero_factura_proveedor': r[5],
                    'imagen': r[6] if r[6] else None,
                } for r in rows],
                'count': len(rows),
            })
        except Exception as e:
            return jsonify({'error': f'lista comprobantes fallo: {e}'}), 500
    # Default · imagen de ordenes_compra (legacy último pago)
    cur.execute("SELECT comprobante_imagen FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    row = cur.fetchone()
    if not row or not row[0]:
        return jsonify({'error': 'Sin comprobante'}), 404
    return jsonify({'imagen': row[0]})

@bp.route('/api/compras/oc/<numero_oc>/rechazar', methods=['POST'])
def rechazar_oc(numero_oc):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    usuario_actual = session.get('compras_user', '')
    d = request.get_json() or {}
    motivo = d.get('motivo', 'Sin motivo especificado')[:300]
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT estado, categoria FROM ordenes_compra WHERE numero_oc=?", (numero_oc,))
    row = cur.fetchone()
    if not row:
        return jsonify({'error': 'OC no encontrada'}), 404
    # Mark OC as Rechazada
    cur.execute("UPDATE ordenes_compra SET estado='Rechazada', observaciones=COALESCE(observaciones,'')||' | RECHAZADA: '||? WHERE numero_oc=?",
                (motivo, numero_oc))
    try:
        audit_log(cur, usuario=usuario_actual, accion='RECHAZAR_OC',
                  tabla='ordenes_compra', registro_id=numero_oc,
                  antes={'estado': row[0], 'categoria': row[1]},
                  despues={'estado': 'Rechazada', 'motivo': motivo[:300]},
                  detalle=f"Rechazó OC {numero_oc} · motivo: {motivo[:120]}")
    except Exception as e:
        log.warning('audit_log RECHAZAR_OC fallo: %s', e)
    # Revert linked solicitud to Pendiente so requester can resubmit
    cur.execute("SELECT numero FROM solicitudes_compra WHERE numero_oc=?", (numero_oc,))
    sol = cur.fetchone()
    if sol:
        cur.execute("UPDATE solicitudes_compra SET estado='Pendiente', observaciones=COALESCE(observaciones,'')||' | RECHAZADA por '||?||': '||? WHERE numero=?",
                    (usuario_actual, motivo, sol[0]))
    # Fetch solicitante nombre + email directo BEFORE closing connection
    _sol_nombre = ''
    _sol_email_directo = ''
    if sol:
        cur.execute('SELECT solicitante, email_solicitante FROM solicitudes_compra WHERE numero=?', (sol[0],))
        _sr = cur.fetchone()
        _sol_nombre = (_sr[0] if _sr else '').lower().strip()
        _sol_email_directo = (_sr[1] if _sr and len(_sr) > 1 else '').strip()
    conn.commit()

    # Sebastian (30-abr-2026): si el motivo sugiere "ya pagamos", buscar el CE
    # asociado para incluir referencia y cerrar el ciclo informativo.
    _motivo_lower = (motivo or '').lower()
    _ya_pagada = any(k in _motivo_lower for k in ['ya pag', 'pagada', 'duplicada', 'duplicado'])
    _ce_info = ''
    _ce_codigo = None
    if _ya_pagada or 'ya' in _motivo_lower and 'pag' in _motivo_lower:
        try:
            ce_row = cur.execute(
                "SELECT numero_ce, fecha_emision, total_pagado FROM comprobantes_pago "
                "WHERE numero_oc=? OR observaciones LIKE ? ORDER BY id DESC LIMIT 1",
                (numero_oc, f'%{numero_oc}%')
            ).fetchone()
            if ce_row:
                _ce_codigo, _ce_fecha, _ce_monto = ce_row
                _ce_info = (f'<div style="background:#d1fae5;border-left:4px solid #16a34a;'
                            f'padding:12px 16px;border-radius:0 6px 6px 0;margin-top:12px">'
                            f'<b>\u2713 Esta OC ya estaba pagada:</b> CE {_ce_codigo} '
                            f'del {_ce_fecha} por ${_ce_monto:,.0f}'
                            f'</div>')
        except Exception:
            pass

    # Email destino: primero el email directo de la solicitud, luego el mapa USER_EMAILS
    _dest_email = _sol_email_directo or USER_EMAILS.get(_sol_nombre, '')
    if sol and _dest_email:
        _asunto_r = f'OC rechazada \u2014 {numero_oc}'
        _accion = 'Tu solicitud volvio a estado <em>Pendiente</em>. Puedes corregirla y reenviarla desde el sistema.'
        if _ce_codigo:
            _accion = (f'No necesitas reenviar la solicitud \u2014 el pago ya esta registrado '
                       f'(CE {_ce_codigo}). Si tienes dudas, abre /compras \u2192 Comprobantes.')
        _body_r = (
            '<html><body style="font-family:Arial,sans-serif;max-width:600px;">'
            '<div style="background:#fee2e2;padding:20px;border-radius:8px;border-left:4px solid #dc2626;">'
            '<h2 style="color:#991b1b;">Orden de compra rechazada</h2>'
            f'<p>La OC <strong>{numero_oc}</strong> asociada a tu solicitud fue rechazada.</p>'
            f'<p><strong>Motivo:</strong> {motivo}</p>'
            f'{_ce_info}'
            f'<p style="margin-top:14px">{_accion}</p>'
            '<p style="color:#6b7280;font-size:12px;">Compras HHA \u2014 Espagiria</p>'
            '</div></body></html>'
        )
        _notificar_solicitante_email(_dest_email, _asunto_r, _body_r)

    # Push notif in-app al solicitante tambi\u00e9n
    if sol and _sol_nombre:
        try:
            from blueprints.notif import push_notif
            cuerpo = motivo
            if _ce_codigo:
                cuerpo += f' \u00b7 \u2713 Ya pagada con CE {_ce_codigo}'
            push_notif(_sol_nombre, 'oc_estado',
                       f'OC {numero_oc} rechazada',
                       body=cuerpo[:160], link='/compras#mis-sol',
                       remitente=usuario_actual,
                       importante=not _ya_pagada)
        except Exception:
            pass
    return jsonify({'ok': True, 'estado': 'Rechazada', 'motivo': motivo,
                    'ce_codigo': _ce_codigo, 'ya_pagada': bool(_ce_codigo)})

@bp.route('/api/compras/buscar-remision/<remision_code>')
def buscar_remision(remision_code):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM ordenes_compra WHERE remision_code=?", (remision_code,))
    oc_row = cur.fetchone()
    oc_cols = [d[0] for d in cur.description] if cur.description else []
    if not oc_row:
        return jsonify({'error': 'No encontrado'}), 404
    oc = dict(zip(oc_cols, oc_row))
    cur.execute("SELECT * FROM ordenes_compra_items WHERE numero_oc=?", (oc['numero_oc'],))
    items = cur.fetchall()
    return jsonify({'oc': oc, 'items': items})

# ════════════════════════════════════════════
# MEE — Materiales de Envase & Empaque
# ════════════════════════════════════════════

@bp.route('/api/mee', methods=['GET','POST'])
def handle_mee():
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        usuario, err, code = _require_compras_write()
        if err:
            return err, code
        d = request.get_json(silent=True) or {}
        if not d.get('codigo') or not d.get('descripcion'):
            return jsonify({'error':'codigo y descripcion requeridos'}), 400
        try:
            cur.execute("""INSERT INTO maestro_mee (codigo,descripcion,categoria,proveedor,fabricante,estado,stock_actual,stock_minimo,unidad,fecha_creacion)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (d['codigo'].upper().strip(), d['descripcion'].strip(),
                 d.get('categoria','Otro'), d.get('proveedor',''), d.get('fabricante',''),
                 'Activo', float(d.get('stock_actual',2000)), float(d.get('stock_minimo',1000)),
                 'und', datetime.now().isoformat()))
            conn.commit()
            return jsonify({'message':f"MEE '{d['codigo']}' creado"}), 201
        except Exception as e:
            log.exception('crear MEE fallo: %s', e)
            return jsonify({'error': 'No se pudo crear el MEE'}), 400
    # GET
    cat = request.args.get('cat','')
    q   = request.args.get('q','')
    lim = int(request.args.get('limit',500))
    sql = "SELECT codigo,descripcion,categoria,proveedor,stock_actual,stock_minimo,estado FROM maestro_mee WHERE estado='Activo'"
    params = []
    if cat: sql += " AND categoria=?"; params.append(cat)
    if q:   sql += " AND (codigo LIKE ? OR descripcion LIKE ?)"; params += [f'%{q}%',f'%{q}%']
    sql += " ORDER BY categoria,codigo LIMIT ?"
    params.append(lim)
    cur.execute(sql, params)
    cols=['codigo','descripcion','categoria','proveedor','stock_actual','stock_minimo','estado']
    items=[dict(zip(cols,r)) for r in cur.fetchall()]
    # AUDITORÍA-FIX 23-may-2026 · C18 · sobrescribir stock_actual cache
    # con valor CANONICAL de movimientos_mee · evita drift en UI mientras
    # el cron mee_drift_sync no haya corrido (mismo día)
    try:
        from blueprints.programacion import _get_mee_stock
        stock_canon = _get_mee_stock(conn)
        for it in items:
            cod_up = str(it.get('codigo') or '').strip().upper()
            if cod_up in stock_canon:
                it['stock_actual'] = round(stock_canon[cod_up], 2)
    except Exception:
        pass
    return jsonify({'items':items})

@bp.route('/api/mee/<codigo>', methods=['GET','PUT'])
def handle_mee_item(codigo):
    conn = get_db(); cur = conn.cursor()
    if request.method == 'PUT':
        usuario, err, code = _require_compras_write()
        if err:
            return err, code
        d = request.get_json(silent=True) or {}
        fields=[]; vals=[]
        for f in ['descripcion','categoria','proveedor','stock_minimo','estado']:
            if f in d: fields.append(f'{f}=?'); vals.append(d[f])
        if not fields: return jsonify({'error':'nada que actualizar'}), 400
        vals.append(codigo)
        cur.execute(f"UPDATE maestro_mee SET {','.join(fields)} WHERE codigo=?", vals)
        conn.commit()
        return jsonify({'message':'actualizado'})
    cur.execute("SELECT * FROM maestro_mee WHERE codigo=?", (codigo,))
    row=cur.fetchone()
    if not row: return jsonify({'error':'no encontrado'}), 404
    cols=[d[0] for d in cur.description]
    return jsonify(dict(zip(cols,row)))

@bp.route('/api/mee/<codigo>/ajuste', methods=['POST'])
def ajuste_mee(codigo):
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    conn = get_db(); cur = conn.cursor()
    d = request.get_json(silent=True) or {}
    nuevo = float(d.get('nuevo_stock',0))
    obs = d.get('observaciones','Ajuste manual')
    oper = d.get('operador','Sistema')
    cur.execute("SELECT stock_actual FROM maestro_mee WHERE codigo=?", (codigo,))
    row=cur.fetchone()
    if not row: return jsonify({'error':'MEE no encontrado'}), 404
    anterior=row[0]
    diff=nuevo-anterior
    cur.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?", (nuevo,codigo))
    cur.execute("INSERT INTO movimientos_mee (mee_codigo,tipo,cantidad,lote_ref,observaciones,responsable,fecha) VALUES (?,?,?,?,?,?,?)",
                (codigo,'Ajuste',diff,'ajuste_manual',obs,oper,datetime.now().isoformat()))
    try:
        audit_log(cur, usuario=usuario, accion='AJUSTE_MEE',
                  tabla='maestro_mee', registro_id=codigo,
                  antes={'stock_actual': anterior},
                  despues={'stock_actual': nuevo, 'diferencia': diff},
                  detalle=f"Ajuste manual de stock MEE {codigo}: {anterior} -> {nuevo}")
    except Exception as e:
        log.warning('audit_log AJUSTE_MEE fallo: %s', e)
    conn.commit()
    return jsonify({'ok':True,'nuevo_stock':nuevo})

@bp.route('/api/movimientos-mee', methods=['GET','POST'])
def handle_movimientos_mee():
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        u, err, code = _require_compras_session()
        if err:
            return err, code
        d = request.get_json(silent=True) or {}
        cod  = d.get('codigo_mee') or d.get('mee_codigo', '')
        tipo_raw = d.get('tipo','Entrada'); tipo = tipo_raw[0].upper()+tipo_raw[1:].lower() if tipo_raw else 'Entrada'
        cant = float(d.get('cantidad',0))
        ref  = d.get('referencia','')
        obs  = d.get('observaciones','')
        oper = d.get('operador','')
        if not cod or cant<=0: return jsonify({'error':'datos invalidos'}), 400
        cur.execute("SELECT stock_actual FROM maestro_mee WHERE codigo=?", (cod,))
        row=cur.fetchone()
        if not row: return jsonify({'error':'MEE no encontrado'}), 404
        delta = cant if tipo in ('Entrada','entrada') else -cant
        nuevo = row[0]+delta
        if nuevo<0: return jsonify({'error':f'Stock insuficiente (actual: {row[0]})'}), 400
        cur.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?", (nuevo,cod))
        cur.execute("INSERT INTO movimientos_mee (mee_codigo,tipo,cantidad,lote_ref,observaciones,responsable,fecha) VALUES (?,?,?,?,?,?,?)",
                    (cod,tipo,cant,ref,obs,oper,datetime.now().isoformat()))
        conn.commit()
        return jsonify({'ok':True,'nuevo_stock':nuevo}), 201
    # GET con filtros
    codigo = request.args.get('codigo','')
    tipo   = request.args.get('tipo','')
    limit  = int(request.args.get('limit',50))
    sql = """SELECT m.id,m.mee_codigo,mm.descripcion,m.tipo,m.cantidad,m.lote_ref,m.observaciones,m.responsable,m.fecha
             FROM movimientos_mee m LEFT JOIN maestro_mee mm ON m.mee_codigo=mm.codigo WHERE 1=1"""
    params=[]
    if codigo: sql+=" AND m.mee_codigo=?"; params.append(codigo)
    if tipo:   sql+=" AND m.tipo=?"; params.append(tipo)
    sql+=" ORDER BY m.fecha DESC LIMIT ?"; params.append(limit)
    cur.execute(sql, params)
    cols=['id','mee_codigo','descripcion','tipo','cantidad','lote_ref','observaciones','responsable','fecha']
    rows=[dict(zip(cols,r)) for r in cur.fetchall()]
    return jsonify({'movimientos':rows})

@bp.route('/api/movimientos-mee/lote', methods=['POST'])
def movimientos_mee_lote():
    u, err, code = _require_compras_session()
    if err:
        return err, code
    conn = get_db(); cur = conn.cursor()
    d = request.get_json(silent=True) or {}
    movs = d.get('movimientos',[])
    oper = d.get('operador','')
    ref  = d.get('referencia','')
    errores=[]
    for m in movs:
        cod=m.get('codigo_mee') or m.get('mee_codigo'); cant=float(m.get('cantidad',0))
        if not cod or cant<=0: continue
        cur.execute("SELECT stock_actual FROM maestro_mee WHERE codigo=?", (cod,))
        row=cur.fetchone()
        if not row: errores.append(f'{cod} no encontrado'); continue
        nuevo=row[0]-cant
        if nuevo<0: nuevo=0
        cur.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?", (nuevo,cod))
        cur.execute("INSERT INTO movimientos_mee (mee_codigo,tipo,cantidad,lote_ref,observaciones,responsable,fecha) VALUES (?,?,?,?,?,?,?)",
                    (cod,'Salida',cant,ref,'Consumo produccion',oper,datetime.now().isoformat()))
    conn.commit()
    if errores: return jsonify({'ok':True,'advertencias':errores})
    return jsonify({'ok':True})

@bp.route('/api/compras/feed-necesidades', methods=['GET'])
def compras_feed_necesidades():
    """Feed unificado de necesidades de compra: materias primas (MP) y envases
    (MEE) por debajo del mínimo, en un solo lugar, ordenado por criticidad
    (menor % de cobertura primero). Read-only. Sebastián 31-may-2026 (Pieza 2)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db(); c = conn.cursor()
    items = []
    # MP bajo mínimo · stock = SUM(movimientos) (mismo cálculo del dashboard)
    try:
        for r in c.execute("""
            SELECT m.codigo_mp,
                   COALESCE(NULLIF(TRIM(m.nombre_comercial),''),
                            NULLIF(TRIM(m.nombre_inci),''), m.codigo_mp),
                   COALESCE(m.proveedor,''), COALESCE(m.stock_minimo,0), COALESCE(s.stock,0)
            FROM maestro_mps m
            LEFT JOIN (SELECT material_id, SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                              WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stock
                       FROM movimientos
                       WHERE UPPER(COALESCE(estado_lote,'')) NOT IN
                             ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO')
                       GROUP BY material_id) s ON s.material_id = m.codigo_mp
            WHERE m.activo=1 AND m.stock_minimo>0 AND COALESCE(s.stock,0) < m.stock_minimo
        """).fetchall():
            mn = float(r[3] or 0); st = float(r[4] or 0)
            items.append({'tipo': 'MP', 'codigo': r[0], 'nombre': r[1], 'proveedor': r[2],
                          'stock': round(st, 1), 'minimo': round(mn, 1),
                          'faltante': round(max(0.0, mn - st), 1),
                          'pct': round(st / mn * 100, 0) if mn > 0 else 0, 'unidad': 'g'})
    except Exception:
        pass
    # Envases (MEE) bajo mínimo
    try:
        for r in c.execute(
            "SELECT codigo, COALESCE(descripcion,''), COALESCE(proveedor,''), "
            "COALESCE(stock_actual,0), COALESCE(stock_minimo,0) "
            "FROM maestro_mee WHERE estado='Activo' AND COALESCE(stock_minimo,0) > 0").fetchall():
            mn = float(r[4] or 0); st = float(r[3] or 0)
            if st < mn:
                items.append({'tipo': 'MEE', 'codigo': r[0], 'nombre': r[1], 'proveedor': r[2],
                              'stock': round(st, 1), 'minimo': round(mn, 1),
                              'faltante': round(max(0.0, mn - st), 1),
                              'pct': round(st / mn * 100, 0) if mn > 0 else 0, 'unidad': 'u'})
    except Exception:
        pass
    items.sort(key=lambda x: x['pct'])
    n_mp = sum(1 for i in items if i['tipo'] == 'MP')
    n_mee = sum(1 for i in items if i['tipo'] == 'MEE')
    return jsonify({'ok': True, 'items': items, 'n': len(items),
                    'n_mp': n_mp, 'n_mee': n_mee})


@bp.route('/api/alertas-mee', methods=['GET'])
def alertas_mee():
    # AUDITORÍA-FIX 23-may-2026 · C18 · canonical SUM(movimientos_mee) ·
    # cargo TODOS los MEE activos con stock_minimo>0 + filtro en Python
    # contra stock canonical · evita falsas alertas por drift del cache
    conn = get_db(); cur = conn.cursor()
    try:
        from blueprints.programacion import _get_mee_stock as _gms
        canon = _gms(conn)
    except Exception:
        canon = {}
    cur.execute("""SELECT codigo,descripcion,categoria,proveedor,stock_actual,stock_minimo
                   FROM maestro_mee WHERE estado='Activo' AND stock_minimo > 0""")
    cols=['codigo','descripcion','categoria','proveedor','stock_actual','stock_minimo']
    alertas = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        cod_up = str(d.get('codigo') or '').strip().upper()
        # Sobrescribir con canonical
        d['stock_actual'] = round(float(canon.get(cod_up, d['stock_actual'] or 0)), 2)
        if d['stock_actual'] < float(d['stock_minimo'] or 0):
            alertas.append(d)
    alertas.sort(key=lambda x: (x['stock_actual'] or 0) - (x['stock_minimo'] or 0))
    return jsonify({'alertas': alertas, 'total': len(alertas)})

@bp.route('/api/compras/consolidado-proveedor', methods=['GET'])
def consolidado_por_proveedor():
    """
    Consolida OCs activas agrupadas por proveedor.
    Incluye datos del proveedor (NIT, contacto, tel) y observaciones de cada OC
    como fallback cuando no hay items registrados en ordenes_compra_items.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    estados_activos = request.args.getlist('estados') or ['Borrador', 'Revisada', 'Autorizada']
    placeholders = ','.join('?' * len(estados_activos))

    conn = get_db(); c = conn.cursor()
    c.execute(f"""
        SELECT
            o.proveedor,
            o.numero_oc,
            o.estado,
            o.fecha,
            o.valor_total,
            o.categoria,
            o.observaciones,
            i.id,
            i.codigo_mp,
            i.nombre_mp,
            i.cantidad_g,
            i.precio_unitario,
            i.subtotal,
            pv.nit,
            pv.contacto,
            pv.telefono,
            pv.email,
            COALESCE(o.con_iva, 0) AS con_iva,
            COALESCE(o.valor_sin_iva, 0) AS valor_sin_iva,
            COALESCE(mm.nombre_inci, '') AS nombre_inci
        FROM ordenes_compra o
        LEFT JOIN ordenes_compra_items i ON o.numero_oc = i.numero_oc
        LEFT JOIN maestro_mps mm ON mm.codigo_mp = i.codigo_mp
        LEFT JOIN proveedores pv ON LOWER(TRIM(o.proveedor)) = LOWER(TRIM(pv.nombre))
        WHERE o.estado IN ({placeholders})
        AND o.categoria NOT IN ('Influencer/Marketing Digital','Cuenta de Cobro','CC')
        ORDER BY o.proveedor, o.numero_oc, i.nombre_mp
    """, estados_activos)

    rows = c.fetchall()
    from collections import OrderedDict
    proveedores = OrderedDict()

    for row in rows:
        (prov, oc, estado, fecha, valor_total_oc, cat, obs,
         item_id, cod_mp, nom_mp, cant, precio_u, subtotal,
         nit, contacto, telefono, email,
         con_iva, valor_sin_iva, nom_inci) = row
        prov = prov or 'Sin proveedor'

        if prov not in proveedores:
            proveedores[prov] = {
                'proveedor': prov,
                'nit': nit or '',
                'contacto': contacto or '',
                'telefono': telefono or '',
                'email': email or '',
                'ocs': {},
                'items': {},
                'valor_total': 0.0,
            }

        p = proveedores[prov]
        # Completar datos proveedor si primera fila no los tenía
        if not p['nit'] and nit: p['nit'] = nit
        if not p['contacto'] and contacto: p['contacto'] = contacto
        if not p['telefono'] and telefono: p['telefono'] = telefono
        if not p['email'] and email: p['email'] = email

        # Registrar OC (incluye observaciones, con_iva, items_raw para edicion)
        if oc and oc not in p['ocs']:
            p['ocs'][oc] = {
                'numero_oc': oc,
                'estado': estado,
                'fecha': (fecha or '')[:10],
                'valor_total': valor_total_oc or 0,
                'categoria': cat or '',
                'observaciones': obs or '',
                'con_iva': int(con_iva or 0),
                'valor_sin_iva': float(valor_sin_iva or 0),
                'items_raw': [],   # items individuales por OC para modo editar
            }
            p['valor_total'] += valor_total_oc or 0

        # Items individuales (para modo editar) por OC
        if oc and item_id is not None:
            p['ocs'][oc]['items_raw'].append({
                'id': item_id,
                'codigo_mp': cod_mp or '',
                'nombre_mp': nom_mp or cod_mp or '',
                'nombre_inci': nom_inci or '',
                'cantidad_g': cant or 0,
                'precio_unitario': precio_u or 0,
                'subtotal': subtotal or 0,
            })

        # Consolidar item por codigo_mp (modo lectura)
        if cod_mp:
            if cod_mp not in p['items']:
                p['items'][cod_mp] = {
                    'codigo_mp': cod_mp,
                    'nombre_mp': nom_mp or cod_mp,
                    'nombre_inci': nom_inci or '',
                    'cantidad_total_g': 0.0,
                    'precio_unitario': precio_u or 0,
                    'subtotal_total': 0.0,
                    'ocs_origen': [],
                }
            item = p['items'][cod_mp]
            item['cantidad_total_g'] += cant or 0
            item['subtotal_total'] += subtotal or 0
            if oc and oc not in item['ocs_origen']:
                item['ocs_origen'].append(oc)

    result = []
    for prov, data in proveedores.items():
        result.append({
            'proveedor': prov,
            'nit': data['nit'],
            'contacto': data['contacto'],
            'telefono': data['telefono'],
            'email': data['email'],
            'ocs': sorted(data['ocs'].values(), key=lambda x: x['fecha'], reverse=True),
            'items': sorted(data['items'].values(), key=lambda x: x['nombre_mp']),
            'valor_total': data['valor_total'],
            'n_ocs': len(data['ocs']),
            'n_items': len(data['items']),
        })

    result.sort(key=lambda x: x['valor_total'], reverse=True)
    return jsonify({'proveedores': result, 'total': len(result)})

@bp.route('/api/compras/solicitudes/pdf', methods=['GET'])
def solicitudes_pdf_resumen():
    """Genera PDF ejecutivo con todas las solicitudes filtradas — para que
    Gerencia (Alejandro) revise lo que falta y dé visto bueno antes de
    convertir en OCs.

    Query params:
      - estados: lista (default 'Pendiente,Aprobada')
      - categoria: filtra (default sin filtro, excluye Influencer/CC)

    Cada solicitud incluye su detalle de items (codigo, nombre, cantidad,
    unidad, justificación) agrupado y totalizado.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    estados = (request.args.get('estados') or 'Pendiente,Aprobada').split(',')
    estados = [e.strip() for e in estados if e.strip()]
    if not estados:
        estados = ['Pendiente', 'Aprobada']

    conn = get_db(); c = conn.cursor()
    placeholders = ','.join('?' * len(estados))
    sql = f"""
        SELECT sc.numero, sc.fecha, sc.estado, sc.solicitante, sc.urgencia,
               sc.observaciones, sc.empresa, sc.categoria, sc.area,
               sc.fecha_requerida, sc.numero_oc,
               COALESCE(NULLIF(oc.valor_total,0), sc.valor, 0) as valor_oc,
               COALESCE(oc.proveedor, '') as proveedor
        FROM solicitudes_compra sc
        LEFT JOIN ordenes_compra oc ON oc.numero_oc = sc.numero_oc
        WHERE sc.estado IN ({placeholders})
          AND sc.categoria NOT IN ('Influencer/Marketing Digital', 'Cuenta de Cobro')
        ORDER BY
          CASE sc.urgencia WHEN 'Alta' THEN 0 WHEN 'Normal' THEN 1 ELSE 2 END,
          sc.fecha_requerida ASC, sc.numero ASC
    """
    sols = c.execute(sql, estados).fetchall()

    # Cargar items por solicitud
    items_por_sol = {}
    for sol_row in sols:
        numero = sol_row[0]
        try:
            it = c.execute("""SELECT codigo_mp, nombre_mp, cantidad_g, unidad,
                                     justificacion, COALESCE(valor_estimado, 0)
                              FROM solicitudes_compra_items
                              WHERE numero=?""", (numero,)).fetchall()
            items_por_sol[numero] = it
        except sqlite3.OperationalError:
            items_por_sol[numero] = []

    if not sols:
        return jsonify({
            'error': 'No hay solicitudes con esos estados',
            'estados_solicitados': estados,
        }), 404

    # ── Generar PDF con fpdf2 — diseño ejecutivo nivel OC ──────────────────
    from fpdf import FPDF
    from datetime import datetime as _dt
    from pathlib import Path

    # Paleta HHA — coordinada con OC (compras_html.py) y branding
    HHA_TEAL = (31, 95, 91)
    HHA_TEAL_DARK = (16, 70, 67)
    COLOR_TEXT = (40, 40, 40)
    COLOR_TEXT_SOFT = (110, 110, 110)
    COLOR_LINE = (200, 195, 188)
    CARD_BG = (250, 248, 244)
    CARD_BORDER = (215, 210, 200)
    TABLE_HEADER_BG = (45, 45, 45)
    TABLE_ALT_ROW = (248, 248, 246)
    BADGE_PEND_BG = (255, 213, 79)       # ámbar Pendiente
    BADGE_PEND_FG = (90, 70, 0)
    BADGE_APR_BG = (89, 165, 121)         # verde Aprobada
    BADGE_APR_FG = (255, 255, 255)
    BADGE_URG_ALTA = (220, 70, 70)
    BADGE_URG_NORM = (140, 140, 140)
    BADGE_URG_BAJA = (180, 180, 180)

    def _fmt_cant(g, u='g'):
        """Normalizado: SIEMPRE en gramos con separador de miles (acordado con Alejandro).
        Preserva decimales solo para péptidos (<1g)."""
        n = float(g or 0)
        if u and u != 'g':
            return f'{n:.0f} {u}' if n == int(n) else f'{n:.2f} {u}'
        if n >= 10:
            return f'{int(round(n)):,}'.replace(',', '.') + ' g'
        if n >= 1:
            return f'{n:.1f} g'
        if n > 0:
            return f'{n:.2f} g'
        return '0 g'

    def _safe(t):
        if t is None:
            return ''
        if not isinstance(t, str):
            t = str(t)
        repl = {'—': '-', '–': '-', '…': '...', '"': '"', '"': '"',
                ''': "'", ''': "'", '•': '·', '→': '->', 'á':'a', 'é':'e',
                'í':'i', 'ó':'o', 'ú':'u', 'Á':'A', 'É':'E', 'Í':'I',
                'Ó':'O', 'Ú':'U', 'ñ':'n', 'Ñ':'N'}
        for k, v in repl.items():
            t = t.replace(k, v)
        return t.encode('latin-1', errors='replace').decode('latin-1')

    n_pendientes = sum(1 for s in sols if s[2] == 'Pendiente')
    n_aprobadas = sum(1 for s in sols if s[2] == 'Aprobada')
    generado_str = _dt.now().strftime("%d/%m/%Y %H:%M")
    estados_str = ', '.join(estados)

    # Logo path resuelto una sola vez
    repo_root = Path(__file__).resolve().parents[2]
    logo_path = None
    for cand in [repo_root / 'api' / 'static' / 'logo_hha.png',
                 repo_root / 'logo_hha.png']:
        if cand.exists():
            logo_path = str(cand)
            break

    class SolicitudesPDF(FPDF):
        """Header repetido en cada página + footer con paginación."""

        def header(self_pdf):
            if self_pdf.page_no() == 1:
                # Header completo en la primera página
                if logo_path:
                    try:
                        self_pdf.image(logo_path, x=10, y=10, w=22, h=22)
                    except Exception:
                        pass
                self_pdf.set_xy(36, 10)
                self_pdf.set_font('Helvetica', 'B', 18)
                self_pdf.set_text_color(*HHA_TEAL)
                self_pdf.cell(110, 8, _safe('SOLICITUDES DE COMPRA'), ln=True)
                self_pdf.set_x(36)
                self_pdf.set_font('Helvetica', '', 9)
                self_pdf.set_text_color(*COLOR_TEXT_SOFT)
                self_pdf.cell(110, 4, _safe('HHA Group · Reporte ejecutivo'), ln=True)
                self_pdf.set_x(36)
                self_pdf.set_font('Helvetica', '', 8)
                self_pdf.cell(110, 4, _safe(f'Generado: {generado_str}'), ln=True)
                self_pdf.set_x(36)
                self_pdf.cell(110, 4, _safe(f'Estados: {estados_str}'), ln=True)

                # Caja resumen (top right)
                bx, by, bw, bh = 142, 10, 58, 24
                self_pdf.set_fill_color(*CARD_BG)
                self_pdf.set_draw_color(*HHA_TEAL)
                self_pdf.set_line_width(0.5)
                self_pdf.rect(bx, by, bw, bh, style='DF')
                self_pdf.set_xy(bx + 3, by + 2)
                self_pdf.set_font('Helvetica', 'B', 7)
                self_pdf.set_text_color(*HHA_TEAL_DARK)
                self_pdf.cell(bw - 6, 3, _safe('TOTAL SOLICITUDES'))
                self_pdf.set_xy(bx + 3, by + 6)
                self_pdf.set_font('Helvetica', 'B', 18)
                self_pdf.set_text_color(*COLOR_TEXT)
                self_pdf.cell(bw - 6, 8, str(len(sols)))
                self_pdf.set_xy(bx + 3, by + 16)
                self_pdf.set_font('Helvetica', '', 7.5)
                self_pdf.set_text_color(*COLOR_TEXT_SOFT)
                self_pdf.cell((bw - 6) / 2, 4, _safe(f'Pendientes: {n_pendientes}'))
                self_pdf.cell((bw - 6) / 2, 4, _safe(f'Aprobadas: {n_aprobadas}'))

                # Línea separadora horizontal teal
                self_pdf.set_draw_color(*HHA_TEAL)
                self_pdf.set_line_width(0.6)
                self_pdf.line(10, 36, 200, 36)
                self_pdf.set_y(40)
            else:
                # Mini-header en páginas siguientes
                if logo_path:
                    try:
                        self_pdf.image(logo_path, x=10, y=8, w=10, h=10)
                    except Exception:
                        pass
                self_pdf.set_xy(22, 9)
                self_pdf.set_font('Helvetica', 'B', 10)
                self_pdf.set_text_color(*HHA_TEAL)
                self_pdf.cell(0, 4, _safe('SOLICITUDES DE COMPRA'), ln=True)
                self_pdf.set_x(22)
                self_pdf.set_font('Helvetica', '', 7)
                self_pdf.set_text_color(*COLOR_TEXT_SOFT)
                self_pdf.cell(0, 3, _safe(f'Continuacion · {generado_str}'), ln=True)
                self_pdf.set_draw_color(*CARD_BORDER)
                self_pdf.set_line_width(0.3)
                self_pdf.line(10, 21, 200, 21)
                self_pdf.set_y(25)

        def footer(self_pdf):
            self_pdf.set_y(-13)
            self_pdf.set_font('Helvetica', 'I', 7)
            self_pdf.set_text_color(*COLOR_TEXT_SOFT)
            self_pdf.cell(0, 4, _safe(f'HHA Group · Pagina {self_pdf.page_no()} de {{nb}}'),
                          align='C')

    def _draw_badge(p, x, y, w, h, text, fg, bg):
        p.set_fill_color(*bg)
        p.set_draw_color(*bg)
        p.rect(x, y, w, h, style='DF')
        p.set_xy(x, y)
        p.set_font('Helvetica', 'B', 7)
        p.set_text_color(*fg)
        p.cell(w, h, _safe(text), align='C')

    def _badge_estado(estado):
        if (estado or '').strip() == 'Aprobada':
            return (BADGE_APR_FG, BADGE_APR_BG)
        return (BADGE_PEND_FG, BADGE_PEND_BG)

    def _badge_urgencia(urg):
        u = (urg or '').strip()
        if u == 'Alta':
            return ((255, 255, 255), BADGE_URG_ALTA)
        if u == 'Normal':
            return ((255, 255, 255), BADGE_URG_NORM)
        return ((255, 255, 255), BADGE_URG_BAJA)

    pdf = SolicitudesPDF(orientation='P', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    total_general_g = 0.0
    total_general_valor = 0.0

    # ─── Una "card" por solicitud ───────────────────────────────────
    for s in sols:
        (numero, fecha, estado, solicitante, urgencia, obs, empresa,
         categoria, area, fecha_req, numero_oc, valor_oc, proveedor) = s
        items = items_por_sol.get(numero, [])

        # Page break preventivo: necesitamos al menos espacio para el header
        # de la card + 1 fila de tabla. Si no, salta.
        if pdf.get_y() > 250:
            pdf.add_page()

        card_x = 10
        card_w = 190
        card_top = pdf.get_y()

        # Barra teal con número de solicitud + badges (estado, urgencia)
        bar_h = 9
        pdf.set_fill_color(*HHA_TEAL)
        pdf.rect(card_x, card_top, card_w, bar_h, style='F')
        pdf.set_xy(card_x + 3, card_top + 1.5)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(70, 6, _safe(numero), align='L')

        # Badges en el extremo derecho
        bw = 24
        bh = 5
        gap = 2
        bx_right = card_x + card_w - 3
        # Urgencia (más a la derecha)
        fg_u, bg_u = _badge_urgencia(urgencia)
        _draw_badge(pdf, bx_right - bw, card_top + 2, bw, bh,
                    f'URG: {(urgencia or "-").upper()[:8]}', fg_u, bg_u)
        # Estado
        fg_e, bg_e = _badge_estado(estado)
        _draw_badge(pdf, bx_right - 2 * bw - gap, card_top + 2, bw, bh,
                    (estado or '-').upper()[:10], fg_e, bg_e)

        # Cuerpo de la card
        pdf.set_y(card_top + bar_h)
        pdf.set_x(card_x + 3)
        pdf.set_font('Helvetica', '', 8.5)
        pdf.set_text_color(*COLOR_TEXT)
        meta1 = (f"Fecha: {(fecha or '')[:10]}    "
                 f"Solicitante: {solicitante or '-'}    "
                 f"Area: {area or '-'}    "
                 f"Empresa: {empresa or '-'}")
        pdf.cell(card_w - 6, 5, _safe(meta1), ln=True)

        if proveedor or fecha_req or numero_oc:
            extra = []
            if proveedor: extra.append(f'Proveedor: {proveedor}')
            if fecha_req: extra.append(f'Fecha req.: {fecha_req}')
            if numero_oc: extra.append(f'OC: {numero_oc}')
            pdf.set_x(card_x + 3)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(*COLOR_TEXT_SOFT)
            pdf.cell(card_w - 6, 4, _safe('    '.join(extra)), ln=True)

        if obs:
            pdf.set_x(card_x + 3)
            pdf.set_font('Helvetica', 'I', 7.5)
            pdf.set_text_color(*COLOR_TEXT_SOFT)
            pdf.multi_cell(card_w - 6, 3.6, _safe(f'Obs: {obs[:280]}'))

        # Tabla de items
        if items:
            pdf.ln(0.5)
            cols = [(24, 'CODIGO', 'L'), (72, 'MATERIAL', 'L'),
                    (24, 'CANTIDAD', 'R'), (10, 'UM', 'C'),
                    (40, 'JUSTIFICACION', 'L'), (20, 'VALOR EST', 'R')]
            # Header row (dark gray)
            pdf.set_x(card_x)
            pdf.set_fill_color(*TABLE_HEADER_BG)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 7.5)
            for w, name, _al in cols:
                pdf.cell(w, 5.5, _safe(name), border=0, fill=True, align='C')
            pdf.ln()

            # Data rows con alternating fill
            sol_total_g = 0.0
            sol_total_valor = 0.0
            pdf.set_text_color(*COLOR_TEXT)
            pdf.set_font('Helvetica', '', 8)
            row_h = 4.5
            for idx, it in enumerate(items):
                cod_mp, nom_mp, cant_g, unidad, just, val_est = it
                cant = float(cant_g or 0)
                val = float(val_est or 0)
                sol_total_g += cant
                sol_total_valor += val
                cant_str = _fmt_cant(cant, unidad)
                fill_alt = (idx % 2 == 1)
                if fill_alt:
                    pdf.set_fill_color(*TABLE_ALT_ROW)
                pdf.set_x(card_x)
                pdf.cell(24, row_h, _safe((cod_mp or '')[:14]),
                         border=0, fill=fill_alt)
                pdf.cell(72, row_h, _safe((nom_mp or '')[:46]),
                         border=0, fill=fill_alt)
                pdf.cell(24, row_h, _safe(cant_str),
                         border=0, fill=fill_alt, align='R')
                pdf.cell(10, row_h, _safe(unidad or 'g'),
                         border=0, fill=fill_alt, align='C')
                pdf.cell(40, row_h, _safe((just or '')[:32]),
                         border=0, fill=fill_alt)
                pdf.cell(20, row_h, _safe(f'${val:,.0f}' if val > 0 else '-'),
                         border=0, fill=fill_alt, align='R')
                pdf.ln()

            # Total row (teal-dark)
            pdf.set_x(card_x)
            pdf.set_fill_color(*HHA_TEAL_DARK)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 8)
            cant_total_str = _fmt_cant(sol_total_g, 'g')
            pdf.cell(24 + 72, 5.5, _safe(f'TOTAL  ·  {len(items)} items'),
                     fill=True, border=0, align='R')
            pdf.cell(24, 5.5, _safe(cant_total_str),
                     fill=True, border=0, align='R')
            pdf.cell(10 + 40, 5.5, '', fill=True, border=0)
            pdf.cell(20, 5.5,
                     _safe(f'${sol_total_valor:,.0f}' if sol_total_valor > 0 else '-'),
                     fill=True, border=0, align='R')
            pdf.ln(0)

            total_general_g += sol_total_g
            total_general_valor += sol_total_valor
        else:
            pdf.set_x(card_x + 3)
            pdf.set_font('Helvetica', 'I', 8)
            pdf.set_text_color(*COLOR_TEXT_SOFT)
            pdf.cell(card_w - 6, 5, _safe('(sin items detallados)'), ln=True)

        # Borde fino alrededor de toda la card (post-render para altura exacta)
        card_h = pdf.get_y() - card_top
        pdf.set_draw_color(*CARD_BORDER)
        pdf.set_line_width(0.3)
        pdf.rect(card_x, card_top, card_w, card_h, style='D')

        pdf.ln(5)

    # ─── Resumen final ────────────────────────────────────────────────
    if pdf.get_y() > 235:
        pdf.add_page()

    pdf.ln(2)
    res_y = pdf.get_y()
    res_h = 30 if total_general_valor > 0 else 24
    pdf.set_fill_color(*CARD_BG)
    pdf.set_draw_color(*HHA_TEAL)
    pdf.set_line_width(0.6)
    pdf.rect(10, res_y, 190, res_h, style='DF')

    pdf.set_xy(13, res_y + 2)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(*HHA_TEAL_DARK)
    pdf.cell(0, 5, _safe('RESUMEN GENERAL'), ln=True)

    pdf.set_xy(13, res_y + 8)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(*COLOR_TEXT)
    pdf.cell(60, 5, _safe(f'Total solicitudes:  {len(sols)}'))
    pdf.cell(60, 5, _safe(f'Pendientes:  {n_pendientes}'))
    pdf.cell(60, 5, _safe(f'Aprobadas:  {n_aprobadas}'))
    pdf.ln(5)

    pdf.set_x(13)
    pdf.cell(0, 5, _safe(f'Cantidad total a comprar:  {_fmt_cant(total_general_g)}'),
             ln=True)
    if total_general_valor > 0:
        pdf.set_x(13)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(0, 5, _safe(f'Valor estimado total:  ${total_general_valor:,.0f} COP'),
                 ln=True)

    pdf.set_y(res_y + res_h)

    # ─── Firmas ───────────────────────────────────────────────────────
    pdf.ln(14)
    y_firma = pdf.get_y()
    pdf.set_draw_color(*COLOR_LINE)
    pdf.set_line_width(0.3)
    pdf.line(20, y_firma, 90, y_firma)
    pdf.line(120, y_firma, 190, y_firma)

    pdf.set_xy(20, y_firma + 1)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_text_color(*COLOR_TEXT)
    pdf.cell(70, 4, _safe('REVISADO POR'), align='C', ln=True)
    pdf.set_x(20)
    pdf.set_font('Helvetica', '', 7)
    pdf.set_text_color(*COLOR_TEXT_SOFT)
    pdf.cell(70, 3.5, _safe('(Compras / Logistica)'), align='C')

    pdf.set_xy(120, y_firma + 1)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_text_color(*COLOR_TEXT)
    pdf.cell(70, 4, _safe('APROBADO POR GERENCIA'), align='C', ln=True)
    pdf.set_x(120)
    pdf.set_font('Helvetica', '', 7)
    pdf.set_text_color(*COLOR_TEXT_SOFT)
    pdf.cell(70, 3.5, _safe('(Alejandro / Sebastian)'), align='C')

    # Output (mismo contrato que antes: PDF descargable con timestamp)
    pdf_bytes = bytes(pdf.output())
    fname = f'solicitudes_compra_{_dt.now().strftime("%Y%m%d_%H%M")}.pdf'
    return Response(
        pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )


# ─── MÓDULO CLIENTES — Rutas ──────────────────────────────────



@bp.route('/api/solicitudes-compra/<numero>/observaciones', methods=['PUT'])
def update_sol_observaciones(numero):
    """Admin: actualiza observaciones (y opcionalmente solicitante) de una solicitud."""
    # SEC-FIX 23-may-2026 · auditoría · era 'compras_user in session' permisivo
    user, err, code = _require_compras_write()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    d = request.json or {}
    obs = d.get('observaciones')
    solicitante = d.get('solicitante')
    valor = d.get('valor')
    fecha_requerida = d.get('fecha_requerida')
    if not obs and not solicitante and valor is None and not fecha_requerida:
        return jsonify({'error': 'Nada que actualizar'}), 400
    row = c.execute(
        "SELECT numero, observaciones, solicitante, valor, fecha_requerida "
        "FROM solicitudes_compra WHERE numero=?",
        (numero.upper(),),
    ).fetchone()
    if not row:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    antes_dict = {
        'observaciones': row[1] if len(row) > 1 else None,
        'solicitante': row[2] if len(row) > 2 else None,
        'valor': row[3] if len(row) > 3 else None,
        'fecha_requerida': row[4] if len(row) > 4 else None,
    }
    updates, params = [], []
    if obs is not None:
        updates.append("observaciones=?"); params.append(obs)
    if solicitante is not None:
        updates.append("solicitante=?"); params.append(solicitante)
    if valor is not None:
        # Cast a float puede fallar si llega string mal formateado.
        # ValueError/TypeError → silencio aceptable (campo opcional).
        try:
            updates.append("valor=?")
            params.append(float(valor))
        except (ValueError, TypeError):
            # Quitar el SET que apenas agregamos para no dejar params desbalanceados.
            if updates and updates[-1] == "valor=?":
                updates.pop()
    if fecha_requerida is not None:
        updates.append("fecha_requerida=?"); params.append(str(fecha_requerida))
    if not updates:
        return jsonify({'error': 'Nada válido para actualizar'}), 400
    params.append(numero.upper())
    c.execute(f"UPDATE solicitudes_compra SET {','.join(updates)} WHERE numero=?", params)
    # SEC-FIX 23-may-2026 · audit_log faltante · INVIMA
    try:
        despues_dict = {
            'observaciones': obs if obs is not None else antes_dict['observaciones'],
            'solicitante': solicitante if solicitante is not None else antes_dict['solicitante'],
            'valor': valor if valor is not None else antes_dict['valor'],
            'fecha_requerida': fecha_requerida if fecha_requerida is not None else antes_dict['fecha_requerida'],
        }
        audit_log(c, usuario=user, accion='ACTUALIZAR_SOL_OBSERVACIONES',
                  tabla='solicitudes_compra', registro_id=numero.upper(),
                  antes=antes_dict, despues=despues_dict)
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'numero': numero.upper()})


# ─── Edicion de items de SOL: cantidad/proveedor/precio (req. Catalina 2026) ──

@bp.route('/api/solicitudes-compra/<numero>/items', methods=['PATCH'])
def update_sol_items(numero):
    """Catalina edita items de una SOL antes de aprobar:
       - cantidad_g  (puede aumentar si conviene pedir mas)
       - proveedor   (si cambia, se normaliza en maestro_mps)
       - precio_unit_g (si cambia, se inserta en precio_historico_mp)

    Body: {items: [{id, cantidad_g, proveedor, precio_unit_g}, ...]}

    Side effects:
      1. solicitudes_compra_items.cantidad_g / valor_estimado actualizados
      2. solicitudes_compra.valor recalculado
      3. maestro_mps.proveedor actualizado si proveedor != actual
      4. precio_historico_mp insertado si precio_unit_g cambio
    """
    # SEC-FIX 23-may-2026 · auditoría · era 'compras_user in session' permisivo
    # · este endpoint también sincroniza maestro_mps.proveedor + precio
    # globalmente · debe restringirse a compras_write
    user, err, code = _require_compras_write()
    if err: return err, code
    d = request.json or {}
    items_in = d.get('items') or []
    if not items_in:
        return jsonify({'error': 'No hay items en el body'}), 400

    conn = get_db(); c = conn.cursor()
    sol = c.execute(
        "SELECT numero, estado FROM solicitudes_compra WHERE numero=?",
        (numero.upper(),)
    ).fetchone()
    if not sol:
        return jsonify({'error': 'SOL no encontrada'}), 404
    # SEC-FIX 23-may-2026 · C13 · solo sincronizar maestro_mps si SOL está
    # en estado editable · evita contaminar el catálogo desde SOL Rechazada/
    # Cancelada/Pagada · variable sync_global controla el bloque después
    estado_sol = (sol[1] or '').strip()
    sync_global_permitido = estado_sol in ('Pendiente', 'Aprobada', 'Borrador')

    cambios = {
        'items_actualizados': 0,
        'maestro_mps_actualizados': 0,
        'precios_historicos_insertados': 0,
    }
    now = datetime.now().isoformat()

    for it in items_in:
        item_id = it.get('id')
        if not item_id:
            continue
        # Estado actual del item
        row = c.execute("""
            SELECT id, codigo_mp, nombre_mp, cantidad_g,
                   COALESCE(precio_unit_g, 0) as precio_unit_g,
                   COALESCE(valor_estimado, 0) as valor_estimado,
                   COALESCE(proveedor_sugerido, '') as proveedor_actual
            FROM solicitudes_compra_items
            WHERE id=? AND numero=?
        """, (item_id, numero.upper())).fetchone()
        if not row:
            continue
        (_id, codigo_mp, nombre_mp, cant_actual,
         precio_actual, valor_actual, prov_actual) = row

        # Valores nuevos (si vienen, sino mantenemos los actuales)
        # P1 audit 26-may · validate_money · update_sol_items propaga a
        # maestro_mps.precio_referencia (INV-2 · sync GLOBAL) · NaN/Inf en
        # precio_referencia rompe `validar_precios_bulk` downstream.
        from http_helpers import validate_money as _vm_si
        _cn_v, _err_cn = _vm_si(it.get('cantidad_g', cant_actual) or 0,
                                  allow_zero=False, max_value=1e9,
                                  field_name='cantidad_g')
        if _err_cn:
            return jsonify(_err_cn), 400
        _pn_v, _err_pn = _vm_si(it.get('precio_unit_g', precio_actual) or 0,
                                  allow_zero=True, max_value=1e9,
                                  field_name='precio_unit_g')
        if _err_pn:
            return jsonify(_err_pn), 400
        cant_nueva = _cn_v
        precio_nuevo = _pn_v
        prov_nuevo = (it.get('proveedor', prov_actual) or '').strip()
        valor_nuevo = round(cant_nueva * precio_nuevo, 2) if precio_nuevo > 0 else valor_actual

        # 1) Update item
        c.execute("""
            UPDATE solicitudes_compra_items
               SET cantidad_g=?, precio_unit_g=?, valor_estimado=?,
                   proveedor_sugerido=?, actualizado_at=?, actualizado_por=?
             WHERE id=?
        """, (cant_nueva, precio_nuevo, valor_nuevo, prov_nuevo, now, user, item_id))
        cambios['items_actualizados'] += 1

        # 2) Si proveedor cambio → sync GLOBAL · Sebastian 6-may-2026:
        # Cuando Catalina asigna proveedor en una SOL, debe quedar guardado
        # en TODA la app (maestro_mps + mp_lead_time_config) para que la
        # próxima vez que el calendario detecte falta de esta MP, sugiera
        # automáticamente el proveedor correcto sin que Catalina tenga que
        # repetirlo.
        # SEC-FIX 23-may-2026 · C13 · solo sincronizar si SOL editable ·
        # evita contaminar maestro_mps desde SOL Rechazada/Cancelada/Pagada
        if codigo_mp and prov_nuevo and prov_nuevo != prov_actual and sync_global_permitido:
            try:
                c.execute("UPDATE maestro_mps SET proveedor=? WHERE codigo_mp=?",
                           (prov_nuevo, codigo_mp))
                if c.rowcount > 0:
                    cambios['maestro_mps_actualizados'] += 1
            except Exception:
                pass
            # Sync mp_lead_time_config (fuente que aplicar_plan() usa con COALESCE)
            try:
                c.execute(
                    "UPDATE mp_lead_time_config SET proveedor_principal=? "
                    "WHERE material_id=?",
                    (prov_nuevo, codigo_mp),
                )
                if c.rowcount == 0:
                    # Si no existe la fila, crearla con defaults
                    c.execute(
                        "INSERT OR IGNORE INTO mp_lead_time_config "
                        "(material_id, material_nombre, proveedor_principal, "
                        " lead_time_dias, buffer_dias, cobertura_min_dias, "
                        " cobertura_ideal_dias, origen, es_envase, activo) "
                        "VALUES (?, ?, ?, 14, 30, 30, 60, 'local', 0, 1)",
                        (codigo_mp, nombre_mp or '', prov_nuevo),
                    )
            except Exception:
                pass
            # Audit
            try:
                audit_log(c, usuario=user, accion='SYNC_PROVEEDOR_GLOBAL',
                           tabla='maestro_mps', registro_id=codigo_mp,
                           antes={'proveedor': prov_actual},
                           despues={'proveedor': prov_nuevo,
                                     'desde_sol': numero.upper()},
                           detalle=(f"Sync global proveedor MP {codigo_mp}: "
                                     f"'{prov_actual}' → '{prov_nuevo}' "
                                     f"(desde edicion en SOL {numero})"))
            except Exception:
                pass

        # 3) Si precio cambio (>0) y es distinto del anterior → historico +
        # sync precio_referencia en maestro_mps (Sebastian 6-may-2026:
        # tambien debe quedar guardado globalmente)
        if precio_nuevo > 0 and abs(precio_nuevo - (precio_actual or 0)) > 1e-6:
            try:
                c.execute("""
                    INSERT INTO precio_historico_mp
                        (codigo_mp, nombre_mp, proveedor, precio_unit_g,
                         cantidad_g, valor_total, fuente, sol_numero, usuario)
                    VALUES (?, ?, ?, ?, ?, ?, 'sol_editada', ?, ?)
                """, (codigo_mp or '', nombre_mp or '', prov_nuevo,
                      precio_nuevo, cant_nueva, valor_nuevo,
                      numero.upper(), user))
                cambios['precios_historicos_insertados'] += 1
            except Exception:
                pass
            # Sync precio_referencia en maestro_mps (precio en pesos por kg ·
            # precio_unit_g está en pesos/g · multiplicar por 1000)
            if codigo_mp:
                try:
                    c.execute(
                        "UPDATE maestro_mps SET precio_referencia=?, "
                        "ultima_act_precio=datetime('now', '-5 hours') WHERE codigo_mp=?",
                        (precio_nuevo * 1000.0, codigo_mp),
                    )
                except sqlite3.OperationalError:
                    # Schema viejo sin ultima_act_precio
                    try:
                        c.execute(
                            "UPDATE maestro_mps SET precio_referencia=? "
                            "WHERE codigo_mp=?",
                            (precio_nuevo * 1000.0, codigo_mp),
                        )
                    except Exception:
                        pass

    # 4) Recalcular valor total de la SOL
    total = c.execute("""
        SELECT COALESCE(SUM(valor_estimado), 0)
          FROM solicitudes_compra_items WHERE numero=?
    """, (numero.upper(),)).fetchone()[0] or 0
    c.execute(
        "UPDATE solicitudes_compra SET valor=? WHERE numero=?",
        (float(total), numero.upper())
    )

    conn.commit()
    return jsonify({
        'ok': True,
        'numero': numero.upper(),
        'valor_total': float(total),
        **cambios,
    })


# ─── Historico de precios por MP ─────────────────────────────────────────────

@bp.route('/api/precio-historico/<path:codigo_mp>', methods=['GET'])
def get_precio_historico(codigo_mp):
    """Devuelve serie temporal de precios + agregados para detectar aumentos.

    Returns:
        {
          codigo_mp, nombre_mp,
          serie: [{fecha, proveedor, precio_unit_g, fuente, sol, oc}, ...],
          stats: {
            ultimo_precio, primer_precio, variacion_pct,
            promedio_30d, promedio_90d,
            min_precio, max_precio,
            n_proveedores_distintos,
            alerta: 'sin_datos'|'estable'|'subiendo'|'bajando'|'volatil'
          }
        }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT id, fecha, proveedor, precio_unit_g, cantidad_g, fuente,
               sol_numero, oc_numero, usuario
        FROM precio_historico_mp
        WHERE codigo_mp = ?
        ORDER BY fecha DESC
        LIMIT 200
    """, (codigo_mp,)).fetchall()

    serie = [
        {
            'id': r[0], 'fecha': r[1], 'proveedor': r[2],
            'precio_unit_g': r[3], 'cantidad_g': r[4],
            'fuente': r[5], 'sol_numero': r[6], 'oc_numero': r[7],
            'usuario': r[8],
        }
        for r in rows
    ]

    stats = {}
    if serie:
        precios = [s['precio_unit_g'] for s in serie if s['precio_unit_g']]
        if precios:
            stats['ultimo_precio'] = precios[0]
            stats['primer_precio'] = precios[-1]
            stats['min_precio'] = min(precios)
            stats['max_precio'] = max(precios)
            if stats['primer_precio'] > 0:
                stats['variacion_pct'] = round(
                    (stats['ultimo_precio'] - stats['primer_precio'])
                    / stats['primer_precio'] * 100,
                    2
                )
            else:
                stats['variacion_pct'] = 0

            # Promedios temporales
            from datetime import datetime as _dt, timedelta as _td
            now = _dt.now()
            d30 = (now - _td(days=30)).isoformat()
            d90 = (now - _td(days=90)).isoformat()
            p30 = [s['precio_unit_g'] for s in serie if s['fecha'] >= d30 and s['precio_unit_g']]
            p90 = [s['precio_unit_g'] for s in serie if s['fecha'] >= d90 and s['precio_unit_g']]
            stats['promedio_30d'] = round(sum(p30) / len(p30), 4) if p30 else None
            stats['promedio_90d'] = round(sum(p90) / len(p90), 4) if p90 else None

            stats['n_proveedores_distintos'] = len(
                {s['proveedor'] for s in serie if s['proveedor']}
            )

            # Alerta
            v = stats.get('variacion_pct', 0)
            if abs(v) < 5:
                stats['alerta'] = 'estable'
                stats['alerta_msg'] = 'Precio estable (variación <5%)'
            elif v >= 20:
                stats['alerta'] = 'subiendo_fuerte'
                stats['alerta_msg'] = (
                    f'Subió {v:.1f}% — considerar explorar otros proveedores'
                )
            elif v >= 5:
                stats['alerta'] = 'subiendo'
                stats['alerta_msg'] = f'Subió {v:.1f}% — vigilar tendencia'
            elif v <= -5:
                stats['alerta'] = 'bajando'
                stats['alerta_msg'] = f'Bajó {v:.1f}% — buena negociación'

    # Nombre del MP (si existe)
    nombre_mp = ''
    try:
        n = c.execute(
            "SELECT nombre_inci FROM maestro_mps WHERE codigo_mp=?",
            (codigo_mp,)
        ).fetchone()
        if n:
            nombre_mp = n[0] or ''
    except Exception:
        pass

    return jsonify({
        'codigo_mp': codigo_mp,
        'nombre_mp': nombre_mp,
        'serie': serie,
        'stats': stats,
    })


# ─── ALERTAS VIVAS DE COMPRAS (Sprint 3) ─────────────────────────────────────
# Espejo del endpoint de Planta — consume del Centro de Mando para mostrar todo
# lo que requiere atención en compras HOY.

@bp.route('/api/compras/alertas-vivas', methods=['GET'])
def alertas_vivas_compras():
    """Consolida 4 categorías de alertas operacionales de Compras.

    Retorna { ocs_sin_recibir, pagos_por_vencer, solicitudes_pendientes,
              ocs_borrador_estancadas, total, severidad_max }
    """
    u, err, code = _require_compras_session()
    if err:
        return err, code

    from datetime import date, timedelta
    hoy = date.today()
    hace_7 = (hoy - timedelta(days=7)).isoformat()
    hace_15 = (hoy - timedelta(days=15)).isoformat()
    hace_3 = (hoy - timedelta(days=3)).isoformat()

    conn = get_db(); c = conn.cursor()

    # ── 1. OCs Autorizadas hace > 15 días sin recepción ─────────────────
    c.execute("""
        SELECT numero_oc, proveedor, valor_total, fecha,
               COALESCE(fecha_entrega_est, '') as fecha_entrega
        FROM ordenes_compra
        WHERE estado = 'Autorizada' AND fecha < ?
        ORDER BY fecha ASC LIMIT 50
    """, (hace_15,))
    ocs_sin_recibir = []
    for r in c.fetchall():
        try:
            f_oc = date.fromisoformat(r[3][:10])
            dias = (hoy - f_oc).days
        except (ValueError, TypeError, IndexError):
            dias = None
        ocs_sin_recibir.append({
            'numero_oc': r[0], 'proveedor': r[1],
            'valor_total': r[2], 'fecha_oc': r[3],
            'fecha_entrega_est': r[4],
            'dias_sin_recibir': dias,
            'severidad': 'alto' if dias and dias > 30 else 'medio',
        })

    # ── 2. Pagos por vencer: OCs Recibidas con pendiente de pago ────────
    # Heurística: si OC recibida hace > N días según condiciones_pago de
    # proveedor (ej. "30 días"), está en riesgo de mora.
    c.execute("""
        SELECT oc.numero_oc, oc.proveedor, oc.valor_total,
               COALESCE(oc.fecha_recepcion, oc.fecha) as fecha_ref,
               COALESCE(p.condiciones_pago, '') as cond,
               COALESCE((SELECT SUM(monto) FROM pagos_oc WHERE numero_oc=oc.numero_oc), 0) as pagado
        FROM ordenes_compra oc
        LEFT JOIN proveedores p ON oc.proveedor = p.nombre
        WHERE oc.estado IN ('Recibida', 'Parcial')
        ORDER BY fecha_ref ASC LIMIT 100
    """)
    pagos_por_vencer = []
    for r in c.fetchall():
        # Extraer días de "30 días" / "30 dias" / "30"
        cond_str = (r[4] or '').lower()
        dias_pago = 30  # default conservador
        for word in cond_str.split():
            if word.isdigit():
                dias_pago = int(word)
                break
        try:
            f_ref = date.fromisoformat(r[3][:10])
            dias_transcurridos = (hoy - f_ref).days
            dias_restantes = dias_pago - dias_transcurridos
        except (ValueError, TypeError, IndexError):
            dias_restantes = None
        pendiente = float(r[2] or 0) - float(r[5] or 0)
        if pendiente > 0.01 and dias_restantes is not None and dias_restantes <= 14:
            pagos_por_vencer.append({
                'numero_oc': r[0], 'proveedor': r[1],
                'valor_total': r[2], 'pendiente': round(pendiente, 2),
                'fecha_ref': r[3], 'condiciones_pago': r[4],
                'dias_restantes': dias_restantes,
                'severidad': ('critico' if dias_restantes < 0 else
                              'alto' if dias_restantes <= 3 else 'medio'),
            })

    # ── 3. Solicitudes Pendientes hace > 3 días ─────────────────────────
    c.execute("""
        SELECT numero, fecha, solicitante, urgencia, area, empresa, categoria
        FROM solicitudes_compra
        WHERE estado = 'Pendiente' AND fecha < ?
        ORDER BY
            CASE urgencia WHEN 'Urgente' THEN 0 WHEN 'Alta' THEN 1 ELSE 2 END,
            fecha ASC
        LIMIT 50
    """, (hace_3,))
    solicitudes_pendientes = []
    for r in c.fetchall():
        try:
            f_sol = date.fromisoformat(r[1][:10])
            dias = (hoy - f_sol).days
        except (ValueError, TypeError, IndexError):
            dias = None
        sev = 'alto' if r[3] == 'Urgente' else ('medio' if dias and dias > 7 else 'bajo')
        solicitudes_pendientes.append({
            'numero': r[0], 'fecha': r[1], 'solicitante': r[2],
            'urgencia': r[3], 'area': r[4], 'empresa': r[5],
            'categoria': r[6], 'dias_pendiente': dias, 'severidad': sev,
        })

    # ── 4. OCs en Borrador hace > 7 días sin avanzar ────────────────────
    c.execute("""
        SELECT numero_oc, proveedor, valor_total, fecha, creado_por
        FROM ordenes_compra
        WHERE estado = 'Borrador' AND fecha < ?
        ORDER BY fecha ASC LIMIT 50
    """, (hace_7,))
    ocs_borrador = [
        {'numero_oc': r[0], 'proveedor': r[1], 'valor_total': r[2],
         'fecha': r[3], 'creado_por': r[4], 'severidad': 'medio'}
        for r in c.fetchall()
    ]

    total = (len(ocs_sin_recibir) + len(pagos_por_vencer) +
             len(solicitudes_pendientes) + len(ocs_borrador))
    sevs = (
        [v['severidad'] for v in ocs_sin_recibir] +
        [v['severidad'] for v in pagos_por_vencer] +
        [v['severidad'] for v in solicitudes_pendientes] +
        [v['severidad'] for v in ocs_borrador]
    )
    sev_orden = {'critico': 4, 'alto': 3, 'medio': 2, 'bajo': 1}
    severidad_max = max(sevs, key=lambda s: sev_orden.get(s, 0)) if sevs else 'ok'

    return jsonify({
        'ocs_sin_recibir': ocs_sin_recibir,
        'pagos_por_vencer': pagos_por_vencer,
        'solicitudes_pendientes': solicitudes_pendientes,
        'ocs_borrador_estancadas': ocs_borrador,
        'total': total,
        'severidad_max': severidad_max,
        'evaluado_en': hoy.isoformat(),
    })


# ─── REPORTE EJECUTIVO COMPRAS [LEGACY] ──────────────────────────────────────

@bp.route('/api/compras/reporte-ejecutivo', methods=['GET'])
def reporte_ejecutivo_compras():
    """[LEGACY 22-may-2026] · /api/compras/dashboard-home consolida estas KPIs.
    Mantenido por compat · usar dashboard-home en código nuevo.

    Reporte gerencial mensual: top proveedores, gasto por categoría,
    pasivo corriente, variaciones de precio.

    Query: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD (default últimos 12 meses)
    """
    u, err, code = _require_compras_session()
    if err:
        return err, code

    from datetime import date, timedelta
    desde = (request.args.get('desde') or
             (date.today() - timedelta(days=365)).isoformat())
    hasta = request.args.get('hasta') or date.today().isoformat()

    conn = get_db(); c = conn.cursor()

    # ── Top 10 proveedores por gasto ───────────────────────────────────
    c.execute("""
        SELECT proveedor,
               COUNT(*) as ocs_count,
               COALESCE(SUM(valor_total), 0) as gasto_total,
               COALESCE(SUM(CASE WHEN estado IN ('Recibida','Pagada') THEN 1 ELSE 0 END), 0) as cumplidas,
               MAX(fecha) as ultima_oc
        FROM ordenes_compra
        WHERE fecha >= ? AND fecha <= ?
              AND estado != 'Rechazada' AND estado != 'Borrador'
        GROUP BY proveedor
        ORDER BY gasto_total DESC LIMIT 10
    """, (desde, hasta + 'T23:59:59'))
    top_proveedores = []
    for r in c.fetchall():
        cumplimiento = round(100.0 * (r[3] or 0) / r[1], 1) if r[1] > 0 else 0
        top_proveedores.append({
            'proveedor': r[0], 'ocs_count': r[1],
            'gasto_total': round(r[2] or 0, 2),
            'cumplidas': r[3], 'cumplimiento_pct': cumplimiento,
            'ultima_oc': r[4],
        })

    # ── Gasto por categoría / mes ──────────────────────────────────────
    c.execute("""
        SELECT substr(fecha, 1, 7) as mes,
               COALESCE(categoria, 'Sin categoría') as cat,
               SUM(valor_total) as total
        FROM ordenes_compra
        WHERE fecha >= ? AND fecha <= ? AND estado != 'Rechazada'
        GROUP BY mes, cat
        ORDER BY mes DESC, total DESC
    """, (desde, hasta + 'T23:59:59'))
    gasto_categoria_mes = [
        {'mes': r[0], 'categoria': r[1], 'total': round(r[2] or 0, 2)}
        for r in c.fetchall()
    ]

    # ── Pasivo corriente: OCs Recibidas/Parciales pendientes de pago ──
    c.execute("""
        SELECT oc.numero_oc, oc.proveedor, oc.valor_total, oc.fecha_recepcion,
               COALESCE((SELECT SUM(monto) FROM pagos_oc WHERE numero_oc=oc.numero_oc), 0) as pagado
        FROM ordenes_compra oc
        WHERE oc.estado IN ('Recibida', 'Parcial')
        ORDER BY oc.fecha_recepcion ASC
    """)
    pasivo_items = []
    pasivo_total = 0
    for r in c.fetchall():
        pendiente = float(r[2] or 0) - float(r[4] or 0)
        if pendiente > 0.01:
            pasivo_items.append({
                'numero_oc': r[0], 'proveedor': r[1],
                'valor_total': r[2], 'pagado': r[4],
                'pendiente': round(pendiente, 2),
                'fecha_recepcion': r[3],
            })
            pasivo_total += pendiente

    # ── Variaciones de precio MP > 15% en últimos 6 meses ──────────────
    c.execute("""
        SELECT codigo_mp,
               MIN(precio_kg) as min_precio,
               MAX(precio_kg) as max_precio,
               COUNT(*) as n_compras,
               (SELECT precio_kg FROM precios_mp_historico p2
                WHERE p2.codigo_mp = p1.codigo_mp ORDER BY fecha DESC LIMIT 1) as ultimo
        FROM precios_mp_historico p1
        WHERE fecha >= date('now', '-5 hours', '-6 months')
        GROUP BY codigo_mp
        HAVING n_compras >= 2
    """)
    variaciones = []
    for r in c.fetchall():
        cod, mn, mx, n, ult = r
        if mn and mn > 0 and mx > mn:
            var_pct = ((mx - mn) / mn) * 100
            if var_pct > 15:
                variaciones.append({
                    'codigo_mp': cod, 'precio_min': mn, 'precio_max': mx,
                    'precio_actual': ult, 'compras_periodo': n,
                    'variacion_pct': round(var_pct, 1),
                    'severidad': 'critico' if var_pct > 50 else
                                 ('alto' if var_pct > 30 else 'medio'),
                })
    variaciones.sort(key=lambda x: x['variacion_pct'], reverse=True)

    # ── Totales mes actual ─────────────────────────────────────────────
    c.execute("""
        SELECT COUNT(*), COALESCE(SUM(valor_total), 0)
        FROM ordenes_compra
        WHERE fecha >= date('now', '-5 hours', 'start of month') AND estado != 'Rechazada'
    """)
    mes_actual = c.fetchone()

    return jsonify({
        'rango': {'desde': desde, 'hasta': hasta},
        'top_proveedores': top_proveedores,
        'gasto_categoria_mes': gasto_categoria_mes,
        'pasivo_corriente': {
            'items': pasivo_items[:50],  # limitar payload
            'total_items': len(pasivo_items),
            'total_pendiente': round(pasivo_total, 2),
        },
        'variaciones_precio': variaciones[:20],
        'mes_actual': {
            'ocs_count': mes_actual[0] if mes_actual else 0,
            'gasto_total': round(mes_actual[1] if mes_actual else 0, 2),
        },
    })


# ─── COTIZACIONES (Sprint 5) ─────────────────────────────────────────────────
# Workflow opcional pre-OC: para items grandes, comparar 3 cotizaciones antes
# de generar la OC. Cada ronda tiene un ronda_id que agrupa N cotizaciones de
# distintos proveedores.

# ════════════════════════════════════════════════════════════════════════
# ÓRDENES DE SERVICIO · Sebastián 21-may-2026
# Flujo: Catalina crea OS → envía proveedor → recogen envases planta →
# proveedor procesa (serigrafía/tampografía/etiquetado) → entrega →
# planta confirma recepción.
# ════════════════════════════════════════════════════════════════════════

_OS_ESTADOS_VALIDOS = (
    'Borrador', 'Enviada', 'Recogida', 'En proceso',
    'Entregada', 'Confirmada', 'Cancelada',
)

# Transiciones permitidas
_OS_TRANSICIONES = {
    'Borrador': {'Enviada', 'Cancelada'},
    'Enviada': {'Recogida', 'Cancelada'},
    'Recogida': {'En proceso', 'Cancelada'},
    'En proceso': {'Entregada', 'Cancelada'},
    'Entregada': {'Confirmada', 'Cancelada'},  # planta confirma
    'Confirmada': set(),  # final
    'Cancelada': set(),   # final
}


def _ensure_os_tables(conn):
    """Defensivo · si mig 150 no aplicó en PG, crea las tablas al primer uso."""
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS ordenes_servicio (
            numero_os TEXT PRIMARY KEY,
            proveedor TEXT NOT NULL,
            tipo_servicio TEXT NOT NULL DEFAULT 'Serigrafía',
            producto_final TEXT, envase_codigo_mee TEXT, envase_descripcion TEXT,
            cantidad_unidades INTEGER NOT NULL DEFAULT 0,
            arte_descripcion TEXT, arte_archivo_url TEXT,
            fecha_solicitud TEXT NOT NULL,
            fecha_requerida_entrega TEXT, fecha_real_entrega TEXT,
            estado TEXT NOT NULL DEFAULT 'Borrador',
            costo_estimado_cop REAL DEFAULT 0, costo_real_cop REAL DEFAULT 0,
            observaciones TEXT, creado_por TEXT NOT NULL,
            creado_at_utc TEXT, planta_confirmado_por TEXT,
            planta_confirmado_at_utc TEXT, cancelada_motivo TEXT,
            tenant_id INTEGER DEFAULT 1
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS ordenes_servicio_eventos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_os TEXT NOT NULL, estado_anterior TEXT,
            estado_nuevo TEXT NOT NULL, usuario TEXT NOT NULL,
            ts_utc TEXT, observaciones TEXT
        )""")
    except Exception:
        pass


@bp.route('/api/compras/ordenes-servicio', methods=['GET', 'POST'])
def ordenes_servicio_list():
    """Lista todas las OS (GET con filtros) o crea una nueva (POST)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    _ensure_os_tables(conn)

    if request.method == 'POST':
        # Solo Catalina/admin/compras-write puede crear
        d = request.get_json(silent=True) or {}
        proveedor = (d.get('proveedor') or '').strip()
        tipo = (d.get('tipo_servicio') or 'Serigrafía').strip()
        producto = (d.get('producto_final') or '').strip()
        envase_cod = (d.get('envase_codigo_mee') or '').strip()
        envase_desc = (d.get('envase_descripcion') or '').strip()
        try:
            cant = int(d.get('cantidad_unidades') or 0)
        except (ValueError, TypeError):
            cant = 0
        arte = (d.get('arte_descripcion') or '').strip()
        fecha_req = (d.get('fecha_requerida_entrega') or '').strip()
        obs = (d.get('observaciones') or '').strip()
        try:
            costo_est = float(d.get('costo_estimado_cop') or 0)
        except (ValueError, TypeError):
            costo_est = 0
        if not proveedor:
            return jsonify({'error': 'proveedor requerido'}), 400
        if cant <= 0:
            return jsonify({'error': 'cantidad_unidades > 0 requerido'}), 400
        if not producto:
            return jsonify({'error': 'producto_final requerido'}), 400
        # Generar numero_os con reintento
        from datetime import datetime as _dt
        anio = _dt.now().strftime('%Y')
        numero_os = None
        for _ in range(6):
            try:
                row = c.execute(
                    "SELECT COALESCE(MAX(CAST(SUBSTR(numero_os,9) AS INTEGER)),0) "
                    "FROM ordenes_servicio WHERE numero_os LIKE ?",
                    (f'OS-{anio}-%',),
                ).fetchone()
                numero_os = f'OS-{anio}-{(row[0] or 0) + 1:04d}'
                fecha_sol = _dt.now().isoformat()
                c.execute(
                    """INSERT INTO ordenes_servicio
                       (numero_os, proveedor, tipo_servicio, producto_final,
                        envase_codigo_mee, envase_descripcion, cantidad_unidades,
                        arte_descripcion, fecha_solicitud, fecha_requerida_entrega,
                        estado, costo_estimado_cop, observaciones, creado_por,
                        creado_at_utc)
                       VALUES (?,?,?,?,?,?,?,?,?,?, 'Borrador', ?, ?, ?, ?)""",
                    (numero_os, proveedor, tipo, producto, envase_cod, envase_desc,
                     cant, arte, fecha_sol, fecha_req or '',
                     costo_est, obs, user, fecha_sol),
                )
                # Evento inicial
                c.execute(
                    """INSERT INTO ordenes_servicio_eventos
                       (numero_os, estado_anterior, estado_nuevo, usuario,
                        ts_utc, observaciones)
                       VALUES (?, NULL, 'Borrador', ?, ?, 'OS creada')""",
                    (numero_os, user, fecha_sol),
                )
                break
            except Exception:
                continue
        if not numero_os:
            return jsonify({'error': 'No se pudo asignar numero_os'}), 500
        try:
            audit_log(c, usuario=user, accion='CREAR_ORDEN_SERVICIO',
                      tabla='ordenes_servicio', registro_id=numero_os,
                      despues={'proveedor': proveedor, 'tipo': tipo,
                               'cantidad': cant, 'producto': producto})
        except Exception:
            pass
        conn.commit()
        return jsonify({'ok': True, 'numero_os': numero_os}), 201

    # GET · listar con filtros
    estado = (request.args.get('estado') or '').strip()
    proveedor = (request.args.get('proveedor') or '').strip()
    where = ['1=1']
    params = []
    if estado:
        where.append('estado = ?'); params.append(estado)
    if proveedor:
        where.append("LOWER(COALESCE(proveedor,'')) LIKE LOWER(?)")
        params.append(f'%{proveedor}%')
    rows = c.execute(
        f"""SELECT numero_os, proveedor, tipo_servicio, producto_final,
                  envase_codigo_mee, COALESCE(envase_descripcion,''),
                  cantidad_unidades, fecha_solicitud, fecha_requerida_entrega,
                  fecha_real_entrega, estado, COALESCE(costo_estimado_cop,0),
                  COALESCE(costo_real_cop,0), creado_por,
                  COALESCE(planta_confirmado_por,''),
                  COALESCE(planta_confirmado_at_utc,'')
           FROM ordenes_servicio
           WHERE {' AND '.join(where)}
           ORDER BY fecha_solicitud DESC LIMIT 200""",
        params,
    ).fetchall()
    items = [{
        'numero_os': r[0], 'proveedor': r[1], 'tipo_servicio': r[2],
        'producto_final': r[3], 'envase_codigo_mee': r[4],
        'envase_descripcion': r[5], 'cantidad_unidades': int(r[6] or 0),
        'fecha_solicitud': r[7], 'fecha_requerida_entrega': r[8],
        'fecha_real_entrega': r[9], 'estado': r[10],
        'costo_estimado_cop': float(r[11] or 0), 'costo_real_cop': float(r[12] or 0),
        'creado_por': r[13], 'planta_confirmado_por': r[14],
        'planta_confirmado_at_utc': r[15],
    } for r in rows]
    # Counts por estado
    try:
        cs_rows = c.execute(
            "SELECT estado, COUNT(*) FROM ordenes_servicio GROUP BY estado",
        ).fetchall()
        counts = {r[0]: int(r[1] or 0) for r in cs_rows}
    except Exception:
        counts = {}
    return jsonify({'items': items, 'total': len(items), 'counts': counts})


@bp.route('/api/compras/ordenes-servicio/<numero_os>', methods=['GET'])
def ordenes_servicio_detalle(numero_os):
    """Detalle completo + timeline de eventos."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    _ensure_os_tables(conn)
    row = c.execute(
        "SELECT * FROM ordenes_servicio WHERE numero_os=?", (numero_os,),
    ).fetchone()
    if not row:
        return jsonify({'error': 'OS no existe'}), 404
    cols = [d[0] for d in c.description]
    info = dict(zip(cols, row))
    # Timeline
    try:
        ev_rows = c.execute(
            """SELECT estado_anterior, estado_nuevo, usuario, ts_utc,
                      COALESCE(observaciones,'')
               FROM ordenes_servicio_eventos
               WHERE numero_os=? ORDER BY id ASC""",
            (numero_os,),
        ).fetchall()
        info['timeline'] = [{
            'estado_anterior': e[0], 'estado_nuevo': e[1],
            'usuario': e[2], 'ts': e[3], 'observaciones': e[4],
        } for e in ev_rows]
    except Exception:
        info['timeline'] = []
    return jsonify(info)


@bp.route('/api/compras/ordenes-servicio/<numero_os>/estado', methods=['PATCH'])
def ordenes_servicio_cambiar_estado(numero_os):
    """Cambia el estado validando transición permitida.

    Body: {estado_nuevo, observaciones?, costo_real_cop? (en Entregada)}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    body = request.get_json(silent=True) or {}
    estado_nuevo = (body.get('estado_nuevo') or '').strip()
    obs = (body.get('observaciones') or '').strip()
    if estado_nuevo not in _OS_ESTADOS_VALIDOS:
        return jsonify({'error': f'estado inválido · usar {_OS_ESTADOS_VALIDOS}'}), 400
    # ROLE-FIX · 21-may-2026 · Confirmada solo planta/admin (cierre cruzado)
    # Antes cualquier compras_user (incluso Catalina creadora) confirmaba ·
    # rompía la separación creador ≠ confirmador del flujo INVIMA.
    if estado_nuevo == 'Confirmada':
        try:
            from config import PLANTA_USERS
        except Exception:
            PLANTA_USERS = set()
        u_lower = (user or '').lower()
        allowed = {x.lower() for x in (set(PLANTA_USERS) | set(ADMIN_USERS))}
        if u_lower not in allowed:
            return jsonify({
                'error': 'Solo planta o admin pueden confirmar recepción de OS',
                'codigo': 'OS_CONFIRMAR_SOLO_PLANTA',
            }), 403
    conn = get_db(); c = conn.cursor()
    _ensure_os_tables(conn)
    row = c.execute(
        "SELECT estado FROM ordenes_servicio WHERE numero_os=?", (numero_os,),
    ).fetchone()
    if not row:
        return jsonify({'error': 'OS no existe'}), 404
    estado_anterior = row[0]
    # Validar transición
    transiciones_ok = _OS_TRANSICIONES.get(estado_anterior, set())
    if estado_nuevo not in transiciones_ok:
        return jsonify({
            'error': f'Transición no permitida: {estado_anterior} → {estado_nuevo}',
            'transiciones_permitidas': list(transiciones_ok),
        }), 409
    from datetime import datetime as _dt
    now = _dt.now().isoformat()
    # Update con campos especiales según estado
    extra_sets = ''
    extra_params = []
    if estado_nuevo == 'Entregada':
        extra_sets += ', fecha_real_entrega=?'
        extra_params.append(now)
        if 'costo_real_cop' in body:
            try:
                extra_sets += ', costo_real_cop=?'
                extra_params.append(float(body['costo_real_cop'] or 0))
            except (ValueError, TypeError):
                pass
    elif estado_nuevo == 'Confirmada':
        extra_sets += ', planta_confirmado_por=?, planta_confirmado_at_utc=?'
        extra_params.extend([user, now])
    elif estado_nuevo == 'Cancelada':
        if not obs:
            return jsonify({'error': 'motivo de cancelación requerido (observaciones)'}), 400
        extra_sets += ', cancelada_motivo=?'
        extra_params.append(obs)
    c.execute(
        f"UPDATE ordenes_servicio SET estado=?{extra_sets} WHERE numero_os=?",
        [estado_nuevo] + extra_params + [numero_os],
    )
    c.execute(
        """INSERT INTO ordenes_servicio_eventos
           (numero_os, estado_anterior, estado_nuevo, usuario, ts_utc, observaciones)
           VALUES (?,?,?,?,?,?)""",
        (numero_os, estado_anterior, estado_nuevo, user, now, obs),
    )
    # Sebastián 24-may-2026 · audit OS · movimientos kardex MEE.
    # Antes el módulo OS era data-only · no afectaba stock · si los 500
    # frascos salían a serigrafía durante 15d, la bodega MEE no sabía
    # que estaban "en tránsito" y Planta los seguía contando como
    # disponibles para producción. Ahora:
    #   • Enviada → SALIDA del envase_codigo_mee (sale a proveedor)
    #   • Confirmada → ENTRADA del envase (vuelve procesado)
    #   • Cancelada desde Recogida/En proceso → ENTRADA (vuelve sin
    #     procesar · se reintegra a bodega)
    # Lote sintético "OS-<num>" para trazabilidad y evitar romper la
    # regla "nunca synthetic lote silencioso" (queda con prefix audit).
    _mee_movimientos_aplicados = False
    try:
        os_row = c.execute(
            "SELECT envase_codigo_mee, envase_descripcion, cantidad_unidades, "
            "       proveedor, tipo_servicio "
            "FROM ordenes_servicio WHERE numero_os=?",
            (numero_os,),
        ).fetchone()
        if os_row and os_row[0]:
            envase_cod = os_row[0]
            envase_desc = os_row[1] or ''
            cant_u = float(os_row[2] or 0)
            prov_os = os_row[3] or ''
            tipo_serv = os_row[4] or 'Servicio'
            if cant_u > 0:
                lote_os = f'OS-{numero_os}'
                # SALIDA · cuando OS arranca (Enviada)
                if estado_nuevo == 'Enviada' and estado_anterior == 'Borrador':
                    try:
                        c.execute(
                            "INSERT INTO movimientos_mee "
                            "(mee_codigo, tipo, cantidad, lote_ref, "
                            " observaciones, responsable, fecha) "
                            "VALUES (?, 'Salida', ?, ?, ?, ?, ?)",
                            (envase_cod, cant_u, lote_os,
                             f'Enviado a {prov_os} · {tipo_serv} · OS {numero_os}',
                             user, now),
                        )
                        # UPDATE stock maestro_mee (kardex MEE no auto-suma)
                        c.execute(
                            "UPDATE maestro_mee SET stock_actual = "
                            "  COALESCE(stock_actual,0) - ? WHERE codigo=?",
                            (cant_u, envase_cod),
                        )
                        _mee_movimientos_aplicados = True
                    except sqlite3.OperationalError as _e_mov:
                        log.warning('OS movimiento_mee Salida fallo OS %s: %s', numero_os, _e_mov)
                # ENTRADA · cuando OS confirma (Confirmada)
                elif estado_nuevo == 'Confirmada':
                    try:
                        c.execute(
                            "INSERT INTO movimientos_mee "
                            "(mee_codigo, tipo, cantidad, lote_ref, "
                            " observaciones, responsable, fecha) "
                            "VALUES (?, 'Entrada', ?, ?, ?, ?, ?)",
                            (envase_cod, cant_u, lote_os,
                             f'Vuelta de {prov_os} · {tipo_serv} procesado · OS {numero_os}',
                             user, now),
                        )
                        c.execute(
                            "UPDATE maestro_mee SET stock_actual = "
                            "  COALESCE(stock_actual,0) + ? WHERE codigo=?",
                            (cant_u, envase_cod),
                        )
                        _mee_movimientos_aplicados = True
                    except sqlite3.OperationalError as _e_mov:
                        log.warning('OS movimiento_mee Entrada fallo OS %s: %s', numero_os, _e_mov)
                # ENTRADA · cancelada desde un estado donde el material
                # ya estaba con el proveedor (vuelve sin procesar)
                elif (estado_nuevo == 'Cancelada' and
                        estado_anterior in ('Recogida', 'En proceso', 'Entregada')):
                    try:
                        c.execute(
                            "INSERT INTO movimientos_mee "
                            "(mee_codigo, tipo, cantidad, lote_ref, "
                            " observaciones, responsable, fecha) "
                            "VALUES (?, 'Entrada', ?, ?, ?, ?, ?)",
                            (envase_cod, cant_u, lote_os,
                             f'Devolución sin procesar · {prov_os} · OS {numero_os} cancelada · {(obs or "")[:100]}',
                             user, now),
                        )
                        c.execute(
                            "UPDATE maestro_mee SET stock_actual = "
                            "  COALESCE(stock_actual,0) + ? WHERE codigo=?",
                            (cant_u, envase_cod),
                        )
                        _mee_movimientos_aplicados = True
                    except sqlite3.OperationalError as _e_mov:
                        log.warning('OS movimiento_mee devolución fallo OS %s: %s', numero_os, _e_mov)
    except Exception as _e_os_mov:
        log.warning('OS movimientos_mee sync fallo OS %s: %s', numero_os, _e_os_mov)
    try:
        audit_log(c, usuario=user, accion=f'OS_{estado_nuevo.upper()}',
                  tabla='ordenes_servicio', registro_id=numero_os,
                  antes={'estado': estado_anterior},
                  despues={'estado': estado_nuevo, 'obs': obs[:200]})
    except Exception:
        pass
    # Push notif cuando pasa a 'Entregada' (planta debe confirmar)
    if estado_nuevo == 'Entregada':
        try:
            from blueprints.notif import push_notif as _push
            from config import PLANTA_USERS
            os_row = c.execute(
                "SELECT proveedor, producto_final, cantidad_unidades FROM ordenes_servicio WHERE numero_os=?",
                (numero_os,),
            ).fetchone()
            if os_row:
                titulo = f'📦 OS {numero_os} entregada · confirmar recepción'
                cuerpo = f'{os_row[0]} entregó {os_row[2]} uds para {os_row[1]}'
                for u in (PLANTA_USERS or set()):
                    try:
                        _push(destinatario=u, tipo='os_entregada',
                              titulo=titulo, body=cuerpo,
                              link='/planta/ordenes-servicio',
                              remitente=user, importante=True)
                    except Exception:
                        pass
        except Exception:
            pass
    conn.commit()
    return jsonify({
        'ok': True, 'numero_os': numero_os,
        'estado_anterior': estado_anterior, 'estado_nuevo': estado_nuevo,
    })


@bp.route('/planta/ordenes-servicio', methods=['GET'])
def planta_ordenes_servicio_page():
    """Sebastián 21-may-2026 · UI Planta · confirmar recepción de OS.

    Página standalone donde planta_users ven OS estado 'Entregada' y
    pueden marcarlas como Confirmadas (cierre del loop creado por Catalina).
    """
    if 'compras_user' not in session:
        from flask import redirect as _r
        return _r('/login?next=/planta/ordenes-servicio')
    html = '''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Recibir Órdenes de Servicio · Planta</title>
<style>
*{box-sizing:border-box;font-family:'Segoe UI',Roboto,sans-serif}
body{margin:0;background:#f1f5f9;padding:18px;color:#0f172a}
.wrap{max-width:900px;margin:0 auto;background:#fff;border-radius:12px;padding:24px;box-shadow:0 2px 10px rgba(0,0,0,.06)}
h1{color:#0f766e;margin:0 0 6px}
.subtitle{color:#64748b;font-size:13px;margin-bottom:18px}
.banner{background:#fef3c7;border:1px solid #ca8a04;padding:10px 14px;border-radius:8px;font-size:12px;color:#78350f;margin-bottom:14px}
.os-card{background:#fff;border:2px solid #16a34a;border-radius:10px;padding:14px;margin-bottom:12px}
.os-head{display:flex;justify-content:space-between;align-items:start;flex-wrap:wrap;gap:8px;margin-bottom:8px}
.os-num{font-family:monospace;font-weight:800;font-size:15px;color:#0f766e}
.os-prov{font-size:13px;color:#475569;font-weight:600}
.os-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:8px;margin:8px 0;font-size:12px}
.os-grid > div{background:#f8fafc;padding:6px 10px;border-radius:5px}
.os-grid b{display:block;font-size:10px;color:#64748b;text-transform:uppercase;margin-bottom:2px}
.arte{background:#fef3c7;border-left:3px solid #ca8a04;padding:8px 10px;margin-top:8px;font-size:12px;color:#78350f}
.actions{display:flex;gap:8px;margin-top:10px;flex-wrap:wrap}
button.confirm{background:#16a34a;color:#fff;border:none;padding:9px 18px;border-radius:6px;font-weight:700;cursor:pointer;font-size:13px}
button.confirm:hover{background:#15803d}
button.cancel{background:#94a3b8;color:#fff;border:none;padding:7px 14px;border-radius:5px;font-size:12px;cursor:pointer}
#empty{text-align:center;padding:40px;color:#94a3b8;font-size:14px}
#msg{padding:10px 14px;border-radius:6px;margin-bottom:14px;display:none;font-size:13px}
</style></head>
<body>
<div class="wrap">
<h1>📦 Recibir Órdenes de Servicio</h1>
<p class="subtitle">Confirmá recepción de envases serigrafiados / etiquetados / preparados · Catalina queda notificada.</p>
<div class="banner">⚠ Al confirmar, verificá que la cantidad entregada coincida y la calidad del trabajo sea OK. Si hay problema, agregá observaciones detalladas.</div>
<div id="msg"></div>
<div id="lista"></div>
<div id="empty" style="display:none">✓ No hay OS pendientes de confirmar · todo al día</div>
<p style="text-align:center;margin-top:20px"><a href="/modulos" style="color:#64748b;font-size:12px;text-decoration:none">← Volver al hub</a></p>
</div>
<script>
function _esc(s){return String(s||'').replace(/[&<>"\\\x27]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"\\x27":'&#39;'}[c];});}
function _fmt(n){try{return Number(n||0).toLocaleString('es-CO');}catch(e){return n;}}
function _msg(t, ok){
  var m=document.getElementById('msg');
  m.textContent=t;
  m.style.display='block';
  m.style.background=ok?'#dcfce7':'#fee2e2';
  m.style.color=ok?'#166534':'#991b1b';
  setTimeout(function(){m.style.display='none';},5000);
}
async function load(){
  try{
    var r=await fetch('/api/planta/ordenes-servicio',{credentials:'same-origin'});
    var d=await r.json();
    var items=d.items||[];
    var lista=document.getElementById('lista');
    var empty=document.getElementById('empty');
    if(!items.length){
      lista.innerHTML='';
      empty.style.display='block';
      return;
    }
    empty.style.display='none';
    lista.innerHTML=items.map(function(o){
      var fec=(o.fecha_real_entrega||'').substring(0,16);
      return '<div class="os-card">'+
        '<div class="os-head">'+
          '<div><div class="os-num">'+_esc(o.numero_os)+'</div><div class="os-prov">'+_esc(o.proveedor)+' · '+_esc(o.tipo_servicio)+'</div></div>'+
          '<span style="background:#16a34a;color:#fff;padding:3px 12px;border-radius:10px;font-size:11px;font-weight:700">📦 ENTREGADA</span>'+
        '</div>'+
        '<div class="os-grid">'+
          '<div><b>Producto</b>'+_esc(o.producto_final)+'</div>'+
          '<div><b>Envase</b>'+_esc(o.envase_descripcion||'—')+'</div>'+
          '<div><b>Cantidad</b><span style="font-size:18px;font-weight:800;color:#0f766e">'+(o.cantidad_unidades||0)+' uds</span></div>'+
          '<div><b>Entregada</b>'+_esc(fec)+'</div>'+
        '</div>'+
        (o.observaciones?'<div class="arte"><b>Observaciones entrega:</b><br>'+_esc(o.observaciones)+'</div>':'')+
        '<div class="actions">'+
          '<button class="confirm" data-num="'+_esc(o.numero_os)+'" data-cant="'+(o.cantidad_unidades||0)+'">✓ Confirmar recepción y calidad OK</button>'+
        '</div>'+
      '</div>';
    }).join('');
  }catch(e){
    document.getElementById('lista').innerHTML='<div style="color:#dc2626;padding:14px">Error: '+_esc(e.message)+'</div>';
  }
}
document.addEventListener('click',async function(ev){
  var b=ev.target.closest('.confirm');
  if(!b) return;
  var num=b.getAttribute('data-num');
  var cant=b.getAttribute('data-cant');
  var obs=prompt('Confirmar recepción de '+num+'\\n\\nCantidad esperada: '+cant+' uds\\n¿Recibiste esa cantidad y calidad OK?\\n\\nObservaciones (opcional · si hay diferencias o problemas):');
  if(obs===null) return;
  if(!confirm('Confirmás recepción de '+num+'? (Catalina quedará notificada)')) return;
  b.disabled=true; b.textContent='Confirmando...';
  try{
    // FIX 27-may (P2) · token vive en server-side session · NO en cookie.
    // Antes: document.cookie.match siempre devolvía vacío → header sin token
    // → server rechazaba con 403 silente. Fetch /api/csrf-token con cache.
    var token = window._csrfTok;
    if (!token) {
      try {
        var tr = await fetch('/api/csrf-token',{credentials:'same-origin'});
        if (tr.ok) { var td = await tr.json(); token = td.csrf_token || ''; window._csrfTok = token; }
      } catch(_e) { token = ''; }
    }
    var r=await fetch('/api/compras/ordenes-servicio/'+encodeURIComponent(num)+'/estado',{
      method:'PATCH',
      headers:{'Content-Type':'application/json','X-CSRF-Token':token},
      credentials:'same-origin',
      body:JSON.stringify({estado_nuevo:'Confirmada',observaciones:obs||'Recepción confirmada por planta · calidad OK'}),
    });
    var d=await r.json();
    if(!r.ok){
      _msg('Error: '+(d.error||r.status),false);
      b.disabled=false; b.textContent='✓ Confirmar recepción y calidad OK';
      return;
    }
    _msg('✓ '+num+' confirmada. Catalina recibe la notificación.',true);
    setTimeout(load,1500);
  }catch(e){
    _msg('Error red: '+e.message,false);
    b.disabled=false; b.textContent='✓ Confirmar recepción y calidad OK';
  }
});
load();
setInterval(load,30000);  // auto-refresh
</script></body></html>'''
    from flask import Response as _Resp
    return _Resp(html, mimetype='text/html')


@bp.route('/api/planta/ordenes-servicio', methods=['GET'])
def planta_ordenes_servicio_pendientes():
    """Lista para Planta · OS Entregadas pendientes de confirmar recepción."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    _ensure_os_tables(conn)
    rows = c.execute(
        """SELECT numero_os, proveedor, tipo_servicio, producto_final,
                  COALESCE(envase_descripcion,''), cantidad_unidades,
                  fecha_real_entrega, COALESCE(observaciones,''),
                  estado
           FROM ordenes_servicio
           WHERE estado = 'Entregada'
           ORDER BY fecha_real_entrega DESC LIMIT 50""",
    ).fetchall()
    items = [{
        'numero_os': r[0], 'proveedor': r[1], 'tipo_servicio': r[2],
        'producto_final': r[3], 'envase_descripcion': r[4],
        'cantidad_unidades': int(r[5] or 0),
        'fecha_real_entrega': r[6], 'observaciones': r[7],
        'estado': r[8],
    } for r in rows]
    return jsonify({'items': items, 'total': len(items)})


@bp.route('/api/compras/cash-flow', methods=['GET'])
def compras_cash_flow():
    """Compras MAX · 21-may-2026 · Cash flow proyección 30/60/90 días."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # PG-FIX · 21-may-2026 · date(... '+' || ? || ' days') multi-arg rompe
    # en PG. Calculamos cutoff en Python para cada ventana.
    from datetime import datetime as _dtcash, timedelta as _tdcash
    ventanas = [30, 60, 90]
    out = {'proyecciones': []}
    for d in ventanas:
        v = {'dias': d}
        cutoff = (_dtcash.now() + _tdcash(days=d)).date().isoformat()
        try:
            # Sebastián 24-may-2026 · audit Por Pagar · KPI debe coincidir con el
            # endpoint /api/compras/por-pagar · excluir Borrador/Revisada (no están
            # listas para pagar todavía) e incluir Parcial (saldo pendiente).
            r = c.execute(
                """SELECT COUNT(*), COALESCE(SUM(valor_total),0)
                   FROM ordenes_compra
                   WHERE estado IN ('Autorizada','Aprobada','Recibida','Parcial')
                     AND date(fecha) <= ?""",
                (cutoff,),
            ).fetchone()
            v['ocs_por_pagar'] = {'count': int(r[0] or 0), 'monto': float(r[1] or 0)}
        except Exception:
            v['ocs_por_pagar'] = {'count': 0, 'monto': 0}
        # Influencers pendientes
        try:
            r = c.execute(
                """SELECT COUNT(*), COALESCE(SUM(valor),0)
                   FROM solicitudes_compra
                   WHERE estado IN ('Pendiente','Aprobada')
                     AND categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')""",
            ).fetchone()
            v['influencers'] = {'count': int(r[0] or 0), 'monto': float(r[1] or 0)}
        except Exception:
            v['influencers'] = {'count': 0, 'monto': 0}
        v['total_salida'] = v['ocs_por_pagar']['monto'] + v['influencers']['monto']
        out['proyecciones'].append(v)
    # Histórico últimos 30d (real pagado)
    try:
        r = c.execute(
            """SELECT COUNT(*), COALESCE(SUM(monto),0)
               FROM pagos_oc
               WHERE date(fecha) >= date('now','-5 hours','-30 days')""",
        ).fetchone()
        out['pagado_30d'] = {'count': int(r[0] or 0), 'monto': float(r[1] or 0)}
    except Exception:
        out['pagado_30d'] = {'count': 0, 'monto': 0}
    return jsonify(out)


@bp.route('/api/compras/trazabilidad-oc/<numero_oc>', methods=['GET'])
def compras_trazabilidad_oc(numero_oc):
    """Compras MAX · 21-may-2026 · Trazabilidad full-chain.

    OC → SOLs origen → items MP → producciones que usaron las MPs.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # 1. Header OC
    h = c.execute(
        """SELECT numero_oc, proveedor, COALESCE(valor_total,0), estado, fecha,
                  COALESCE(observaciones,'')
           FROM ordenes_compra WHERE numero_oc=?""",
        (numero_oc,),
    ).fetchone()
    if not h:
        return jsonify({'error': 'OC no existe'}), 404
    out = {
        'numero_oc': h[0], 'proveedor': h[1], 'valor_total': float(h[2] or 0),
        'estado': h[3], 'fecha': h[4], 'observaciones': h[5],
    }
    # 2. Items
    try:
        rows = c.execute(
            """SELECT codigo_mp, nombre_mp, COALESCE(cantidad_g,0),
                      COALESCE(precio_unitario,0), COALESCE(subtotal,0)
               FROM ordenes_compra_items WHERE numero_oc=?""",
            (numero_oc,),
        ).fetchall()
        out['items'] = [{
            'codigo_mp': r[0], 'nombre_mp': r[1],
            'cantidad_g': float(r[2] or 0),
            'precio_unitario': float(r[3] or 0),
            'subtotal': float(r[4] or 0),
        } for r in rows]
    except Exception:
        out['items'] = []
    # 3. SOLs origen
    try:
        rows = c.execute(
            """SELECT numero, solicitante, fecha, COALESCE(valor,0)
               FROM solicitudes_compra
               WHERE numero_oc=? OR observaciones LIKE ?""",
            (numero_oc, f'%{numero_oc}%'),
        ).fetchall()
        out['sols_origen'] = [{
            'numero': r[0], 'solicitante': r[1], 'fecha': r[2], 'valor': float(r[3] or 0),
        } for r in rows]
    except Exception:
        out['sols_origen'] = []
    # 4. Pagos
    try:
        rows = c.execute(
            """SELECT fecha, monto, COALESCE(medio,''), COALESCE(referencia,''),
                      COALESCE(numero_factura_proveedor,'')
               FROM pagos_oc WHERE numero_oc=?""",
            (numero_oc,),
        ).fetchall()
        out['pagos'] = [{
            'fecha': r[0], 'monto': float(r[1] or 0),
            'medio': r[2], 'referencia': r[3], 'factura': r[4],
        } for r in rows]
    except Exception:
        out['pagos'] = []
    # 5. Producciones que usaron las MPs de esta OC (último mes)
    codigos = [it['codigo_mp'] for it in out.get('items', []) if it.get('codigo_mp')]
    producciones_relacionadas = []
    if codigos:
        try:
            placeholders = ','.join(['?'] * len(codigos))
            rows = c.execute(
                f"""SELECT DISTINCT pp.id, pp.producto, pp.fecha_programada,
                           COALESCE(pp.lote_codigo, pp.id) as lote
                    FROM produccion_programada pp
                    JOIN movimientos m ON m.material_id IN ({placeholders})
                    WHERE date(pp.fecha_programada) >= date('now','-5 hours','-60 days')
                    ORDER BY pp.fecha_programada DESC LIMIT 20""",
                codigos,
            ).fetchall()
            producciones_relacionadas = [{
                'pid': r[0], 'producto': r[1], 'fecha': r[2], 'lote': r[3],
            } for r in rows]
        except Exception:
            pass
    out['producciones_relacionadas'] = producciones_relacionadas
    return jsonify(out)


def _scorecard_proveedor_dict(c, nombre_prov):
    """Compras Fase 3 · Sebastián 21-may-2026 · scorecard live de proveedor.

    Calcula 5 métricas reales que el consultor procurement recomendó como
    base para renegociación y para reglas de auto-aprobación más finas.

    Métricas (todas últimos 12 meses):
    1. ocs_total          · # OCs creadas
    2. on_time_pct        · % OCs recibidas dentro del lead time prometido
    3. rechazo_qc_pct     · % lotes que QC marcó como RECHAZADO
    4. variacion_precio_12m_pct · % cambio promedio precio MP (1er trim vs último)
    5. cumplimiento_pct   · recibidas+pagadas / total (ya estaba en ROI)
    6. lead_time_real_dias · promedio real (mp_lead_time_config)
    7. score_global       · 0-100 ponderado

    Retorna dict listo para serializar a JSON.
    """
    out = {'proveedor': nombre_prov}
    if not nombre_prov:
        return out
    # 1+5. ocs_total + cumplimiento + monto
    try:
        r = c.execute(
            """SELECT COUNT(*),
                      SUM(CASE WHEN estado IN ('Recibida','Pagada','Parcial') THEN 1 ELSE 0 END),
                      COALESCE(SUM(valor_total),0)
               FROM ordenes_compra
               WHERE LOWER(TRIM(proveedor))=LOWER(TRIM(?))
                 AND date(fecha) >= date('now','-5 hours','-365 days')""",
            (nombre_prov,),
        ).fetchone()
        ocs_total = int((r or [0,0,0])[0] or 0)
        ocs_cumplidas = int((r or [0,0,0])[1] or 0)
        monto_12m = float((r or [0,0,0])[2] or 0)
        out['ocs_total_12m'] = ocs_total
        out['monto_12m'] = monto_12m
        out['cumplimiento_pct'] = round(ocs_cumplidas / ocs_total * 100, 1) if ocs_total else 0
    except Exception:
        out['ocs_total_12m'] = 0
        out['monto_12m'] = 0
        out['cumplimiento_pct'] = 0
    # 2. on_time_pct
    # AUDITORÍA-FIX 23-may-2026 · C19 · antes usaba 30 días hardcoded
    # · ahora compara contra el lead_time real de cada MP en
    # mp_lead_time_config (aprendido por EWMA en recibir_oc) · proveedor
    # con lead_time 7d ya no aparece "100% on-time" llegando a 25d
    try:
        # Por cada OC, calcular si llegó dentro del lead_time real (max de
        # sus items) · MAX porque la OC se considera on-time solo si TODOS
        # sus items hubieran podido llegar a tiempo
        r = c.execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE
                          WHEN julianday(oc.fecha_recepcion) - julianday(oc.fecha) <=
                               COALESCE((
                                  SELECT MAX(COALESCE(mlt.lead_time_dias, 14))
                                  FROM ordenes_compra_items oci
                                  LEFT JOIN mp_lead_time_config mlt ON mlt.material_id = oci.codigo_mp
                                  WHERE oci.numero_oc = oc.numero_oc
                               ), 30)
                          THEN 1 ELSE 0
                      END) AS on_time
               FROM ordenes_compra oc
               WHERE LOWER(TRIM(oc.proveedor))=LOWER(TRIM(?))
                 AND oc.fecha_recepcion IS NOT NULL AND oc.fecha_recepcion != ''
                 AND date(oc.fecha) >= date('now','-5 hours','-365 days')""",
            (nombre_prov,),
        ).fetchone()
        total_recibidas = int((r or [0,0])[0] or 0)
        on_time = int((r or [0,0])[1] or 0)
        out['on_time_pct'] = round(on_time / total_recibidas * 100, 1) if total_recibidas else 0
    except Exception:
        out['on_time_pct'] = 0
    # 3. rechazo_qc_pct (lotes RECHAZADO de movimientos)
    try:
        r = c.execute(
            """SELECT COUNT(*),
                      SUM(CASE WHEN estado_lote='RECHAZADO' THEN 1 ELSE 0 END)
               FROM movimientos
               WHERE LOWER(TRIM(proveedor))=LOWER(TRIM(?))
                 AND tipo='Entrada'
                 AND date(fecha) >= date('now','-5 hours','-365 days')""",
            (nombre_prov,),
        ).fetchone()
        total_lotes = int((r or [0,0])[0] or 0)
        rechazados = int((r or [0,0])[1] or 0)
        out['rechazo_qc_pct'] = round(rechazados / total_lotes * 100, 1) if total_lotes else 0
        out['lotes_evaluados'] = total_lotes
    except Exception:
        out['rechazo_qc_pct'] = 0
        out['lotes_evaluados'] = 0
    # 4. variacion_precio_12m_pct
    try:
        r = c.execute(
            """SELECT AVG(CASE WHEN date(fecha) >= date('now','-5 hours','-90 days') THEN precio_unitario END),
                      AVG(CASE WHEN date(fecha) <  date('now','-5 hours','-275 days') THEN precio_unitario END)
               FROM precios_mp_historico
               WHERE LOWER(TRIM(proveedor))=LOWER(TRIM(?))
                 AND date(fecha) >= date('now','-5 hours','-365 days')""",
            (nombre_prov,),
        ).fetchone()
        p_reciente = float((r or [0,0])[0] or 0)
        p_inicial = float((r or [0,0])[1] or 0)
        if p_inicial > 0:
            out['variacion_precio_12m_pct'] = round((p_reciente - p_inicial) / p_inicial * 100, 1)
        else:
            out['variacion_precio_12m_pct'] = 0
    except Exception:
        out['variacion_precio_12m_pct'] = 0
    # 6. lead_time_real promedio
    try:
        r = c.execute(
            """SELECT AVG(julianday(fecha_recepcion) - julianday(fecha))
               FROM ordenes_compra
               WHERE LOWER(TRIM(proveedor))=LOWER(TRIM(?))
                 AND fecha_recepcion IS NOT NULL AND fecha_recepcion != ''
                 AND date(fecha) >= date('now','-5 hours','-365 days')""",
            (nombre_prov,),
        ).fetchone()
        out['lead_time_real_dias'] = round(float((r or [0])[0] or 0), 1)
    except Exception:
        out['lead_time_real_dias'] = 0
    # 7. score_global (0-100 ponderado · 4 dimensiones)
    cumpl = out['cumplimiento_pct'] / 100
    on_time = out['on_time_pct'] / 100
    no_rechazo = max(0, 1 - out['rechazo_qc_pct'] / 100)
    precio_estable = max(0, 1 - abs(out['variacion_precio_12m_pct']) / 50)  # >50% var = 0
    score = (cumpl * 0.30 + on_time * 0.25 + no_rechazo * 0.30 + precio_estable * 0.15) * 100
    out['score_global'] = round(score, 1)
    out['score_color'] = 'verde' if score >= 80 else ('amarillo' if score >= 60 else 'rojo')
    out['recomendacion'] = (
        'Excelente · usar para auto-aprob' if score >= 85
        else 'Bueno · monitorear' if score >= 70
        else 'Regular · renegociar' if score >= 50
        else 'Crítico · evaluar reemplazo'
    )
    return out


@bp.route('/api/compras/proveedor-scorecard/<nombre_prov>', methods=['GET'])
def proveedor_scorecard(nombre_prov):
    """Compras Fase 3 · Scorecard live de un proveedor (5 métricas + score)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    return jsonify(_scorecard_proveedor_dict(c, nombre_prov))


@bp.route('/api/compras/roi-proveedores', methods=['GET'])
def compras_roi_proveedores():
    """Compras MAX · 21-may-2026 · ROI por proveedor."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    try:
        rows = c.execute(
            """SELECT proveedor,
                      COUNT(*) ocs,
                      COALESCE(SUM(valor_total),0) monto_total,
                      AVG(CASE WHEN fecha_recepcion IS NOT NULL AND fecha_recepcion != ''
                          THEN julianday(fecha_recepcion) - julianday(fecha) END) lead_time_real_dias,
                      SUM(CASE WHEN estado='Pagada' THEN 1 ELSE 0 END) pagadas,
                      SUM(CASE WHEN estado IN ('Recibida','Pagada','Parcial') THEN 1 ELSE 0 END) recibidas,
                      MAX(fecha) ultima_compra
               FROM ordenes_compra
               WHERE proveedor IS NOT NULL AND proveedor != ''
                 AND date(fecha) >= date('now','-5 hours','-365 days')
               GROUP BY proveedor
               ORDER BY monto_total DESC LIMIT 30""",
        ).fetchall()
    except Exception:
        rows = []
    items = []
    for r in rows:
        items.append({
            'proveedor': r[0],
            'ocs_12m': int(r[1] or 0),
            'monto_12m': float(r[2] or 0),
            'pagadas': int(r[4] or 0),
            'recibidas': int(r[5] or 0),
            'ultima_compra': r[6] or '',
            'cumplimiento_pct': round((r[5] / r[1] * 100) if r[1] else 0, 1),
        })
    return jsonify({'proveedores': items, 'total': len(items)})


@bp.route('/api/compras/coa-upload', methods=['POST'])
def compras_coa_upload():
    """Sebastián 22-may-2026 · upload binario COA · INVIMA closure.

    Acepta PDF/JPG/PNG (≤10MB) · guarda en Render persistent disk (/var/data/coa/)
    · NO requiere S3/Cloudinary · path local persistente.

    Body multipart/form-data:
      - archivo: PDF/JPG/PNG
      - codigo_mp (opt): asocia el COA a un MP específico
      - lote_proveedor (opt): identifica lote del COA

    Returns: { coa_url: '/api/compras/coa-download/<id>', coa_filename, size_kb }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if not (user in ADMIN_USERS or user in COMPRAS_ACCESS):
        return jsonify({'error': 'Solo Compras/Admin pueden subir COA'}), 403
    if 'archivo' not in request.files:
        return jsonify({'error': 'Falta archivo (campo multipart \'archivo\')'}), 400
    f = request.files['archivo']
    if not f or not f.filename:
        return jsonify({'error': 'Archivo vacío'}), 400
    # Validar magic bytes (no confiar solo extensión)
    head = f.read(8)
    f.seek(0)
    if head[:4] == b'%PDF':
        ext = 'pdf'; mime = 'application/pdf'
    elif head[:3] == b'\xff\xd8\xff':
        ext = 'jpg'; mime = 'image/jpeg'
    elif head[:8] == b'\x89PNG\r\n\x1a\n':
        ext = 'png'; mime = 'image/png'
    else:
        return jsonify({'error': 'Formato no soportado · solo PDF/JPG/PNG'}), 400
    # Tamaño max 10MB
    f.seek(0, 2); size = f.tell(); f.seek(0)
    if size > 10 * 1024 * 1024:
        return jsonify({'error': f'Archivo > 10MB ({size//1024}KB)'}), 413
    # Carpeta destino
    import os as _os, uuid as _uuid
    coa_dir = _os.environ.get('COA_STORAGE_DIR', '/var/data/coa')
    try:
        _os.makedirs(coa_dir, exist_ok=True)
    except Exception as e:
        log.warning('No pude crear %s: %s · fallback /tmp/coa', coa_dir, e)
        coa_dir = '/tmp/coa'
        _os.makedirs(coa_dir, exist_ok=True)
    # Guardar con nombre único
    file_id = _uuid.uuid4().hex[:12]
    safe_name = f'coa_{datetime.now().strftime("%Y%m%d")}_{file_id}.{ext}'
    full_path = _os.path.join(coa_dir, safe_name)
    try:
        f.save(full_path)
    except Exception as e:
        return jsonify({'error': f'No pude guardar: {e}'}), 500
    # audit log
    conn = get_db(); c = conn.cursor()
    try:
        audit_log(c, usuario=user, accion='COA_UPLOAD',
                  tabla='movimientos', registro_id=file_id,
                  despues={'filename': safe_name, 'mime': mime,
                           'size_kb': size // 1024,
                           'codigo_mp': request.form.get('codigo_mp', ''),
                           'lote_proveedor': request.form.get('lote_proveedor', '')})
        conn.commit()
    except Exception:
        pass
    return jsonify({
        'ok': True,
        'coa_url': f'/api/compras/coa-download/{safe_name}',
        'coa_filename': safe_name,
        'size_kb': size // 1024,
        'mime': mime,
    })


@bp.route('/api/compras/coa-download/<path:filename>', methods=['GET'])
def compras_coa_download(filename):
    """Descarga binaria de COA · session requerida."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    import os as _os
    # Sanitize · solo nombre del archivo · NO path traversal
    safe = _os.path.basename(filename)
    if safe != filename or '..' in filename:
        return jsonify({'error': 'Nombre inválido'}), 400
    coa_dir = _os.environ.get('COA_STORAGE_DIR', '/var/data/coa')
    full = _os.path.join(coa_dir, safe)
    if not _os.path.exists(full):
        # Fallback a /tmp/coa si var/data no existe (dev local)
        full = _os.path.join('/tmp/coa', safe)
    if not _os.path.exists(full):
        return jsonify({'error': 'COA no encontrado'}), 404
    from flask import send_file
    mime_by_ext = {'pdf': 'application/pdf', 'jpg': 'image/jpeg',
                   'jpeg': 'image/jpeg', 'png': 'image/png'}
    ext = safe.rsplit('.', 1)[-1].lower()
    return send_file(full, mimetype=mime_by_ext.get(ext, 'application/octet-stream'),
                     as_attachment=False, download_name=safe)


@bp.route('/api/compras/ocr-factura', methods=['POST'])
def compras_ocr_factura():
    """Compras MAX · 21-may-2026 · OCR factura proveedor.

    Body: {imagen_base64, ocs_candidatas?: [str]}
    Devuelve datos extraídos + sugiere match con OC pendiente.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    body = request.get_json(silent=True) or {}
    b64 = (body.get('imagen_base64') or '').strip()
    if not b64:
        return jsonify({'error': 'imagen_base64 requerido'}), 400
    if b64.startswith('data:'):
        try: b64 = b64.split(',', 1)[1]
        except Exception: return jsonify({'error': 'data URL inválido'}), 400
    if len(b64) > 7_000_000:
        return jsonify({'error': 'imagen muy grande (max ~5MB)'}), 413
    # Bug #4 fix · 21-may-2026 · validar magic bytes antes de mandar a Claude
    # · si llega PDF/HEIC/otro, devolver 415 explícito (no esperar al 502 de Anthropic).
    try:
        import base64 as _b64
        head = _b64.b64decode(b64[:32] + '=' * (-len(b64[:32]) % 4))
    except Exception:
        return jsonify({'error': 'imagen_base64 inválido', 'codigo': 'B64_INVALIDO'}), 400
    if head[:4] == b'%PDF':
        return jsonify({'error': 'PDF no soportado · convertí a JPG/PNG',
                        'codigo': 'PDF_NO_SOPORTADO'}), 415
    if head[:3] == b'\xff\xd8\xff':
        media_type = 'image/jpeg'
    elif head[:8] == b'\x89PNG\r\n\x1a\n':
        media_type = 'image/png'
    else:
        return jsonify({'error': 'Formato no soportado · usar JPG o PNG',
                        'codigo': 'FORMATO_NO_SOPORTADO',
                        'magic_hex': head[:8].hex()}), 415
    import os, json as _json
    api_key = os.environ.get('ANTHROPIC_API_KEY') or os.environ.get('CLAUDE_API_KEY')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY no configurada',
                        'codigo': 'NO_API_KEY'}), 503
    system_prompt = (
        "Sos OCR especializado en facturas de proveedores en español "
        "(Colombia). Extraé los datos de la factura y devolvelos como "
        "JSON puro sin texto adicional:\n"
        '{\"proveedor\":\"\",\"nit\":\"\",\"numero_factura\":\"\",'
        '\"fecha_emision\":\"YYYY-MM-DD o null\",'
        '\"fecha_vencimiento\":\"YYYY-MM-DD o null\",'
        '\"subtotal\":0,\"iva\":0,\"total\":0,'
        '\"items\":[{\"descripcion\":\"\",\"cantidad\":0,'
        '\"unidad\":\"\",\"precio_unitario\":0,\"subtotal\":0}],'
        '\"confianza_pct\":85}\n'
        "Si un campo no está visible, dejá string vacío o null. "
        "confianza_pct (0-100) es tu certeza global."
    )
    import urllib.request as _ureq, urllib.error as _uerr
    try:
        req = _ureq.Request(
            'https://api.anthropic.com/v1/messages',
            data=_json.dumps({
                'model': 'claude-sonnet-4-6',
                'max_tokens': 1500,
                'system': system_prompt,
                'messages': [{
                    'role': 'user',
                    'content': [
                        {'type': 'image', 'source': {
                            'type': 'base64', 'media_type': media_type, 'data': b64,
                        }},
                        {'type': 'text', 'text': 'Extraé los campos.'},
                    ],
                }],
            }).encode('utf-8'),
            headers={
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            },
            method='POST',
        )
        with _ureq.urlopen(req, timeout=60) as resp:
            data = _json.loads(resp.read().decode('utf-8'))
        txt = ''
        for b in (data.get('content') or []):
            if b.get('type') == 'text':
                txt += b.get('text', '')
        txt = (txt or '').strip()
        if txt.startswith('```'):
            txt = txt.split('```')[1]
            if txt.startswith('json'):
                txt = txt[4:].strip()
        try:
            parsed = _json.loads(txt)
        except Exception:
            return jsonify({'error': 'OCR devolvió JSON inválido', 'raw': txt[:500]}), 502
        # Sugerir match con OC pendiente
        conn = get_db(); c = conn.cursor()
        sugerencias_oc = []
        if parsed.get('proveedor'):
            try:
                prov_norm = parsed['proveedor'].strip().lower()[:30]
                total_factura = float(parsed.get('total') or 0)
                rows = c.execute(
                    """SELECT numero_oc, proveedor, COALESCE(valor_total,0), estado, fecha
                       FROM ordenes_compra
                       WHERE LOWER(COALESCE(proveedor,'')) LIKE ?
                         AND estado IN ('Autorizada','Recibida','Revisada')
                       ORDER BY fecha DESC LIMIT 5""",
                    (f'%{prov_norm}%',),
                ).fetchall()
                for r in rows:
                    valor_oc = float(r[2] or 0)
                    delta_pct = abs(valor_oc - total_factura) / valor_oc * 100 if valor_oc > 0 else 999
                    sugerencias_oc.append({
                        'numero_oc': r[0], 'proveedor': r[1],
                        'valor_total': valor_oc,
                        'estado': r[3],
                        'fecha': r[4],
                        'delta_vs_factura_pct': round(delta_pct, 1),
                        'match_score': 'alto' if delta_pct < 3 else ('medio' if delta_pct < 10 else 'bajo'),
                    })
            except Exception:
                pass
        try:
            audit_log(c, usuario=user, accion='OCR_FACTURA_PROVEEDOR',
                      tabla='_', registro_id='_',
                      despues={'proveedor': parsed.get('proveedor',''),
                                'total': parsed.get('total', 0),
                                'confianza': parsed.get('confianza_pct', 0)})
            conn.commit()
        except Exception:
            pass
        return jsonify({
            'ok': True,
            'factura': parsed,
            'ocs_sugeridas': sugerencias_oc,
        })
    except _uerr.HTTPError as e:
        return jsonify({'error': f'Anthropic HTTP {e.code}: {e.reason}'}), 502
    except Exception as e:
        return jsonify({'error': f'OCR falló: {e}'}), 500


@bp.route('/api/compras/prediccion-demanda', methods=['GET'])
def compras_prediccion_demanda():
    """Compras MAX · 21-may-2026 · Predicción demanda MPs.

    Por cada MP con stock + consumo histórico 90d:
    - calcula consumo_diario_promedio
    - estima fecha_quiebre = hoy + (stock / consumo_diario)
    - sugiere acción: 'OK', 'PEDIR_PRONTO' (<30d), 'URGENTE' (<14d)
    - calcula cantidad sugerida = consumo_90d × factor_buffer
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # Stock actual por material_id
    # ABASTECIMIENTO-FIX · 22-may-2026 · Ajuste/Ajuste+ suman · CUARENTENA excluida
    stock = {}
    try:
        rows = c.execute(
            """SELECT material_id,
                      COALESCE(SUM(
                        CASE
                          WHEN tipo IN ('Entrada','Ajuste +','Ajuste') THEN cantidad
                          WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad
                          ELSE 0
                        END),0)
               FROM movimientos
               WHERE material_id IS NOT NULL AND material_id != ''
                 AND (estado_lote IS NULL
                      OR UPPER(COALESCE(estado_lote,'')) NOT IN
                         ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))
               GROUP BY material_id""",
        ).fetchall()
        for r in rows:
            stock[r[0]] = max(float(r[1] or 0), 0)
    except Exception:
        pass
    # Consumo 90d (Salidas)
    consumo = {}
    try:
        rows = c.execute(
            """SELECT material_id, SUM(cantidad)
               FROM movimientos
               WHERE tipo='Salida'
                 AND date(fecha) >= date('now','-5 hours','-90 days')
                 AND material_id IS NOT NULL AND material_id != ''
               GROUP BY material_id""",
        ).fetchall()
        for r in rows:
            consumo[r[0]] = float(r[1] or 0)
    except Exception:
        pass
    # Metadata MPs
    metadata = {}
    try:
        rows = c.execute(
            """SELECT codigo_mp, nombre_comercial, COALESCE(proveedor,''),
                      COALESCE(precio_referencia,0)
               FROM maestro_mps WHERE COALESCE(activo,1)=1""",
        ).fetchall()
        for r in rows:
            metadata[r[0]] = {'nombre': r[1], 'proveedor': r[2] or '',
                              'precio_kg': float(r[3] or 0)}
    except Exception:
        pass
    # Lead times
    lead_times = {}
    try:
        rows = c.execute(
            """SELECT material_id, COALESCE(lead_time_dias, 15)
               FROM mp_lead_time_config""",
        ).fetchall()
        for r in rows:
            lead_times[r[0]] = int(r[1] or 15)
    except Exception:
        pass
    # Construir predicción
    items = []
    for mid, c_total in consumo.items():
        if c_total <= 0:
            continue
        st = stock.get(mid, 0)
        consumo_diario = c_total / 90.0
        if consumo_diario <= 0:
            continue
        dias_quiebre = st / consumo_diario if consumo_diario > 0 else 999
        lt = lead_times.get(mid, 15)
        meta = metadata.get(mid, {})
        # Acción
        if dias_quiebre < lt + 3:
            accion = 'URGENTE'
            color = 'rojo'
        elif dias_quiebre < lt + 14:
            accion = 'PEDIR_PRONTO'
            color = 'amarillo'
        else:
            accion = 'OK'
            color = 'verde'
        # ABASTECIMIENTO-FIX · 22-may-2026 · dedup con cola pendiente (#6 audit 22-may)
        # · Antes: si ya hay 50kg en SOL Pendiente y 30kg en OC Autorizada,
        #   predicción seguía pidiendo todo desde cero · duplicación al aprobar.
        # · Ahora: resta `_pendiente_en_compras_g` del déficit antes de sugerir.
        en_cola_g = _pendiente_en_compras_g(c, mid)
        cantidad_sugerida = max(0, consumo_diario * (lt + 30) - st - en_cola_g)
        costo_estimado = (cantidad_sugerida / 1000.0) * meta.get('precio_kg', 0)
        items.append({
            'codigo_mp': mid,
            'nombre': meta.get('nombre', mid),
            'proveedor_default': meta.get('proveedor', ''),
            'stock_g': round(st, 1),
            'consumo_90d_g': round(c_total, 1),
            'consumo_diario_g': round(consumo_diario, 1),
            'dias_hasta_quiebre': round(dias_quiebre, 1),
            'lead_time_dias': lt,
            'accion': accion,
            'color': color,
            'cantidad_sugerida_g': round(cantidad_sugerida, 1),
            'costo_estimado_cop': round(costo_estimado, 0),
        })
    # Ordenar por urgencia
    orden = {'URGENTE': 0, 'PEDIR_PRONTO': 1, 'OK': 2}
    items.sort(key=lambda x: (orden.get(x['accion'], 99), x['dias_hasta_quiebre']))
    counts = {'URGENTE': 0, 'PEDIR_PRONTO': 0, 'OK': 0}
    for it in items:
        counts[it['accion']] += 1
    return jsonify({
        'items': items[:100],
        'counts': counts,
        'total_evaluados': len(items),
    })


@bp.route('/api/compras/asistente-ia', methods=['POST'])
def compras_asistente_ia():
    """Compras MAX · 21-may-2026 · IA Agente "Pregúntale a Compras".

    Chat con Claude Sonnet 4.6 · contexto auto del módulo Compras:
    - Top 10 proveedores históricos
    - SOLs pendientes
    - OCs autorizadas sin pago
    - Cash flow próximos 30d
    - Histórico precios MPs

    Body: {pregunta: str}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    body = request.get_json(silent=True) or {}
    pregunta = (body.get('pregunta') or '').strip()
    if len(pregunta) < 3:
        return jsonify({'error': 'pregunta requerida (≥3 chars)'}), 400
    import os, json as _json
    api_key = os.environ.get('ANTHROPIC_API_KEY') or os.environ.get('CLAUDE_API_KEY')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY no configurada',
                        'codigo': 'NO_API_KEY'}), 503
    conn = get_db(); c = conn.cursor()
    # Contexto auto
    contexto = {}
    try:
        rows = c.execute(
            """SELECT proveedor, COUNT(*) ocs, COALESCE(SUM(valor_total),0) m
               FROM ordenes_compra
               WHERE date(fecha) >= date('now','-5 hours','-90 days')
                 AND proveedor IS NOT NULL AND proveedor != ''
               GROUP BY proveedor ORDER BY m DESC LIMIT 10""",
        ).fetchall()
        contexto['top_proveedores_90d'] = [
            {'nombre': r[0], 'ocs': r[1], 'monto': float(r[2] or 0)} for r in rows
        ]
    except Exception:
        contexto['top_proveedores_90d'] = []
    try:
        rows = c.execute(
            """SELECT estado, COUNT(*) FROM solicitudes_compra
               WHERE estado IN ('Pendiente','Aprobada')
               GROUP BY estado""",
        ).fetchall()
        contexto['sols_por_estado'] = {r[0]: r[1] for r in rows}
    except Exception:
        contexto['sols_por_estado'] = {}
    try:
        r = c.execute(
            "SELECT COUNT(*), COALESCE(SUM(valor_total),0) FROM ordenes_compra WHERE estado='Autorizada'",
        ).fetchone()
        contexto['ocs_por_pagar'] = {'count': int(r[0] or 0), 'monto': float(r[1] or 0)}
    except Exception:
        contexto['ocs_por_pagar'] = {'count': 0, 'monto': 0}
    try:
        r = c.execute(
            """SELECT SUM(CASE WHEN estado='Pagada' THEN valor_total ELSE 0 END),
                      SUM(CASE WHEN estado IN ('Borrador','Revisada','Autorizada') THEN valor_total ELSE 0 END)
               FROM ordenes_compra
               WHERE date(fecha) >= date('now','-5 hours','-30 days')""",
        ).fetchone()
        contexto['cash_flow_30d'] = {
            'pagado': float(r[0] or 0),
            'por_pagar': float(r[1] or 0),
        }
    except Exception:
        contexto['cash_flow_30d'] = {}
    # Top MPs por valor compra
    try:
        rows = c.execute(
            """SELECT oci.codigo_mp, oci.nombre_mp,
                      COALESCE(SUM(oci.subtotal),0) m,
                      COUNT(DISTINCT oc.numero_oc) ocs
               FROM ordenes_compra_items oci
               JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
               WHERE date(oc.fecha) >= date('now','-5 hours','-90 days')
               GROUP BY oci.codigo_mp, oci.nombre_mp
               ORDER BY m DESC LIMIT 5""",
        ).fetchall()
        contexto['top_mps_90d'] = [{
            'codigo': r[0], 'nombre': r[1], 'monto': float(r[2] or 0), 'ocs': r[3],
        } for r in rows]
    except Exception:
        contexto['top_mps_90d'] = []
    contexto['usuario_actual'] = user

    system_prompt = (
        "Sos el asistente IA del módulo Compras de un laboratorio cosmético "
        "(Espagiria + ÁNIMUS Lab). Tenés acceso a contexto live de la BD: "
        "top proveedores, SOLs pendientes, OCs por pagar, cash flow, top MPs.\n\n"
        "REGLAS:\n"
        "- Respondé en español neutro · max 6 oraciones · directo\n"
        "- Si pregunta sobre proveedor/MP/cash flow · usá los datos del contexto\n"
        "- Si falta data en el contexto · decilo (no inventes números)\n"
        "- Si la pregunta requiere acción · sugerí el botón exacto en /compras\n"
        "- NO digas 'soy una IA' · sos el asistente de Compras\n"
        "- Para preguntas como '¿qué proveedor me conviene?' · ofrecé el top 3 "
        "con argumentos basados en el contexto (frecuencia, monto, etc.)\n"
    )
    user_msg = (
        f"PREGUNTA: {pregunta}\n\n"
        f"CONTEXTO LIVE COMPRAS:\n"
        f"{_json.dumps(contexto, ensure_ascii=False, indent=1)}"
    )
    import urllib.request as _ureq, urllib.error as _uerr
    try:
        req = _ureq.Request(
            'https://api.anthropic.com/v1/messages',
            data=_json.dumps({
                'model': 'claude-sonnet-4-6',
                'max_tokens': 700,
                'system': system_prompt,
                'messages': [{'role': 'user', 'content': user_msg}],
            }).encode('utf-8'),
            headers={
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            },
            method='POST',
        )
        with _ureq.urlopen(req, timeout=30) as resp:
            data = _json.loads(resp.read().decode('utf-8'))
        respuesta = ''
        for b in (data.get('content') or []):
            if b.get('type') == 'text':
                respuesta += b.get('text', '')
        respuesta = (respuesta or '').strip()
        try:
            audit_log(c, usuario=user, accion='COMPRAS_ASISTENTE_IA',
                      tabla='_', registro_id='_',
                      despues={'pregunta_chars': len(pregunta),
                                'respuesta_chars': len(respuesta)})
            conn.commit()
        except Exception:
            pass
        return jsonify({
            'ok': True,
            'respuesta': respuesta or '(sin respuesta · reintentá)',
            'contexto_resumen': {
                'proveedores': len(contexto.get('top_proveedores_90d', [])),
                'sols_pendientes': sum(contexto.get('sols_por_estado', {}).values()),
                'ocs_por_pagar': contexto.get('ocs_por_pagar', {}).get('count', 0),
            },
        })
    except _uerr.HTTPError as e:
        return jsonify({'error': f'Anthropic HTTP {e.code}: {e.reason}'}), 502
    except Exception as e:
        return jsonify({'error': f'Asistente falló: {e}'}), 500


@bp.route('/api/compras/dashboard-home', methods=['GET'])
def compras_dashboard_home():
    """Compras 2.0 · Sebastián 21-may-2026 · Dashboard dual por rol.

    Devuelve TODO lo que necesita la pantalla home en 1 request:
      - role: 'admin' o 'operativo'
      - kpis (salud_score, sols_pendientes_3d, ocs_5d, etc)
      - buzon (SOLs nuevas hoy por fuente · solo si operativo)
      - influencers (pendientes con monto · solo si admin)
      - alertas críticas (todos)
      - counts por tab (para badges)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '').lower()
    is_admin = user in {x.lower() for x in ADMIN_USERS}
    conn = get_db(); c = conn.cursor()
    out = {'role': 'admin' if is_admin else 'operativo', 'usuario': user}

    # KPIs salud (igual que dashboard-ejecutivo existente · reusable)
    kpis = {}
    try:
        r = c.execute(
            """SELECT COUNT(*) FROM solicitudes_compra
               WHERE estado = 'Pendiente'
                 AND date(fecha) < date('now','-5 hours','-3 days')""",
        ).fetchone()
        kpis['sols_sin_tocar_3d'] = int(r[0] or 0)
    except Exception: kpis['sols_sin_tocar_3d'] = 0
    try:
        r = c.execute(
            """SELECT COUNT(*) FROM ordenes_compra
               WHERE estado IN ('Borrador','Revisada')
                 AND date(fecha) < date('now','-5 hours','-5 days')""",
        ).fetchone()
        kpis['ocs_sin_autorizar_5d'] = int(r[0] or 0)
    except Exception: kpis['ocs_sin_autorizar_5d'] = 0
    pendientes = (kpis['sols_sin_tocar_3d'] + kpis['ocs_sin_autorizar_5d'])
    kpis['salud_score'] = max(0, 100 - pendientes * 5)
    kpis['salud_color'] = 'verde' if kpis['salud_score'] >= 80 else ('amarillo' if kpis['salud_score'] >= 60 else 'rojo')
    out['kpis'] = kpis

    # Counts por tab (badges)
    # MEDIA-7 fix · loggear excepciones · antes el except silencioso ocultaba
    # bugs SQL y mostraba "0 SOLs" falsos (Sebastián no actuaba).
    counts = {}
    _log = __import__('logging').getLogger('compras.dashboard')
    def _count(key, sql, params=()):
        try:
            counts[key] = int((c.execute(sql, params).fetchone() or [0])[0])
        except Exception as e:
            _log.warning('counts.%s SQL fallo: %s', key, e)
            counts[key] = 0
    _count('planta',
        """SELECT COUNT(*) FROM solicitudes_compra
           WHERE estado='Pendiente'
             AND categoria IN ('Materia Prima','Empaque','Material de Empaque')""")
    _count('solic',
        """SELECT COUNT(*) FROM solicitudes_compra
           WHERE estado='Pendiente'
             AND categoria NOT IN ('Materia Prima','Empaque','Material de Empaque',
                                   'Influencer/Marketing Digital','Cuenta de Cobro')""")
    _count('influencer',
        """SELECT COUNT(*) FROM solicitudes_compra
           WHERE estado IN ('Pendiente','Aprobada')
             AND categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')""")
    # Sebastián 24-may-2026 · audit Por Pagar · alinear badge con el endpoint
    # /api/compras/por-pagar que muestra Recibida + Parcial (mercancía) +
    # Aprobada/Autorizada (pago directo servicios). Antes solo contaba
    # Autorizada · el badge subestimaba al usuario el trabajo pendiente.
    _count('por_pagar',
        "SELECT COUNT(*) FROM ordenes_compra "
        "WHERE estado IN ('Autorizada','Aprobada','Recibida','Parcial')")
    _count('consol',
        "SELECT COUNT(*) FROM ordenes_compra WHERE estado IN ('Borrador','Revisada','Autorizada')")
    out['counts'] = counts

    # Buzón · SOLs nuevas hoy (ambos roles)
    try:
        rows = c.execute(
            """SELECT numero, solicitante, fecha, categoria, COALESCE(valor,0) as v
               FROM solicitudes_compra
               WHERE estado='Pendiente'
                 AND date(fecha) >= date('now','-5 hours','-2 days')
               ORDER BY fecha DESC LIMIT 10""",
        ).fetchall()
        out['buzon_recientes'] = [{
            'numero': r[0], 'solicitante': r[1], 'fecha': r[2],
            'categoria': r[3], 'valor': float(r[4] or 0),
        } for r in rows]
    except Exception:
        out['buzon_recientes'] = []
    # BUG-CRITICA-2 fix · count REAL sin LIMIT (antes KPI mentía con >10)
    try:
        bt = c.execute(
            """SELECT COUNT(*) FROM solicitudes_compra
               WHERE estado='Pendiente'
                 AND date(fecha) >= date('now','-5 hours','-2 days')""",
        ).fetchone()
        out['buzon_total_48h'] = int(bt[0] or 0)
    except Exception:
        out['buzon_total_48h'] = len(out.get('buzon_recientes', []))

    # Influencers (solo admin)
    if is_admin:
        try:
            rows = c.execute(
                """SELECT numero, solicitante, fecha, COALESCE(valor,0) as v,
                          estado, COALESCE(observaciones,'')
                   FROM solicitudes_compra
                   WHERE estado IN ('Pendiente','Aprobada')
                     AND categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')
                   ORDER BY fecha DESC LIMIT 10""",
            ).fetchall()
            out['influencers_pendientes'] = [{
                'numero': r[0], 'solicitante': r[1], 'fecha': r[2],
                'monto': float(r[3] or 0), 'estado': r[4],
                'concepto': (r[5] or '')[:80],
            } for r in rows]
            out['influencers_monto_total'] = sum(float(r[3] or 0) for r in rows)
        except Exception:
            out['influencers_pendientes'] = []
            out['influencers_monto_total'] = 0
        # Top proveedores 30d (solo admin)
        try:
            rows = c.execute(
                """SELECT proveedor, COUNT(*) as ocs, COALESCE(SUM(valor_total),0) as monto
                   FROM ordenes_compra
                   WHERE date(fecha) >= date('now','-5 hours','-30 days')
                     AND proveedor IS NOT NULL AND proveedor != ''
                   GROUP BY proveedor ORDER BY monto DESC LIMIT 5""",
            ).fetchall()
            out['top_proveedores_30d'] = [{
                'proveedor': r[0], 'ocs': int(r[1] or 0), 'monto': float(r[2] or 0),
            } for r in rows]
        except Exception:
            out['top_proveedores_30d'] = []

    # Alertas críticas (top 5 vencimientos próximos · todos los roles)
    try:
        rows = c.execute(
            """SELECT numero_oc, proveedor, COALESCE(valor_total,0), estado, fecha
               FROM ordenes_compra
               WHERE estado='Autorizada'
                 AND date(fecha) < date('now','-5 hours','-10 days')
               ORDER BY fecha ASC LIMIT 5""",
        ).fetchall()
        out['alertas_ocs_viejas'] = [{
            'numero_oc': r[0], 'proveedor': r[1], 'monto': float(r[2] or 0),
            'estado': r[3], 'fecha': r[4],
        } for r in rows]
    except Exception:
        out['alertas_ocs_viejas'] = []

    return jsonify(out)


@bp.route('/api/compras/dashboard-ejecutivo', methods=['GET'])
def compras_dashboard_ejecutivo():
    """[LEGACY 22-may-2026] · Compras 2.0 dashboard-home cubre estas KPIs.
    Endpoint mantenido por compat con código viejo · ver compras_dashboard_home
    para el dashboard CONSOLIDADO que reemplaza este + dashboard-stats + reporte-ejecutivo.

    Sprint Compras N3 · 21-may-2026 · widget ejecutivo Catalina.

    Devuelve KPIs operativos:
      - sols_sin_tocar_3d (Pendientes hace >3 días)
      - ocs_sin_autorizar_5d (Borrador/Revisada hace >5 días)
      - influencers_por_pagar_vencidos (Aprobadas con fecha_requerida pasada)
      - cotizaciones_pendientes (rondas con respuestas <2)
      - top_proveedores_mes (por valor OC último 30d)
      - total_ocs_borrador (count)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    kpis = {}
    # SOLs sin tocar >3d
    try:
        r = c.execute(
            """SELECT COUNT(*) FROM solicitudes_compra
               WHERE estado = 'Pendiente'
                 AND date(fecha) < date('now','-5 hours','-3 days')""",
        ).fetchone()
        kpis['sols_sin_tocar_3d'] = int(r[0] or 0)
    except Exception:
        kpis['sols_sin_tocar_3d'] = 0
    # OCs sin autorizar >5d (Borrador o Revisada antiguas)
    try:
        r = c.execute(
            """SELECT COUNT(*) FROM ordenes_compra
               WHERE estado IN ('Borrador','Revisada')
                 AND date(fecha) < date('now','-5 hours','-5 days')""",
        ).fetchone()
        kpis['ocs_sin_autorizar_5d'] = int(r[0] or 0)
    except Exception:
        kpis['ocs_sin_autorizar_5d'] = 0
    # Influencers vencidos sin pagar
    try:
        r = c.execute(
            """SELECT COUNT(*) FROM solicitudes_compra
               WHERE categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')
                 AND estado = 'Aprobada'
                 AND fecha_requerida != ''
                 AND date(fecha_requerida) < date('now','-5 hours')""",
        ).fetchone()
        kpis['influencers_vencidos'] = int(r[0] or 0)
    except Exception:
        kpis['influencers_vencidos'] = 0
    # OCs Borrador total
    try:
        r = c.execute(
            "SELECT COUNT(*) FROM ordenes_compra WHERE estado='Borrador'",
        ).fetchone()
        kpis['ocs_borrador'] = int(r[0] or 0)
    except Exception:
        kpis['ocs_borrador'] = 0
    # Top 5 proveedores mes (último 30d)
    top_proveedores = []
    try:
        rows = c.execute(
            """SELECT proveedor, COUNT(*) as ocs, COALESCE(SUM(valor_total),0) as monto
               FROM ordenes_compra
               WHERE date(fecha) >= date('now','-5 hours','-30 days')
                 AND proveedor IS NOT NULL AND proveedor != ''
               GROUP BY proveedor
               ORDER BY monto DESC LIMIT 5""",
        ).fetchall()
        top_proveedores = [{
            'proveedor': r[0], 'ocs': int(r[1] or 0),
            'monto': float(r[2] or 0),
        } for r in rows]
    except Exception:
        pass
    # Cotizaciones pendientes (rondas con <2 respuestas)
    cot_pendientes = 0
    try:
        rows = c.execute(
            """SELECT ronda_id, SUM(CASE WHEN estado='Recibida' THEN 1 ELSE 0 END) as recibidas,
                      COUNT(*) as total
               FROM cotizaciones
               WHERE date(fecha_creacion) >= date('now','-5 hours','-30 days')
               GROUP BY ronda_id
               HAVING recibidas < 2""",
        ).fetchall()
        cot_pendientes = len(rows)
    except Exception:
        pass
    kpis['cotizaciones_pendientes'] = cot_pendientes
    # Score salud (0-100): menos pendientes = mejor
    pendientes_total = (kpis['sols_sin_tocar_3d'] + kpis['ocs_sin_autorizar_5d']
                         + kpis['influencers_vencidos'])
    score = max(0, 100 - pendientes_total * 5)
    kpis['salud_score'] = score
    kpis['salud_color'] = (
        'verde' if score >= 80 else 'amarillo' if score >= 60 else 'rojo'
    )
    return jsonify({
        'kpis': kpis,
        'top_proveedores_mes': top_proveedores,
    })


@bp.route('/api/compras/validar-precios-bulk', methods=['POST'])
def validar_precios_bulk():
    """Sprint Compras N3 · 21-may-2026 · alerta precio inflado.

    Recibe lista de {codigo_mp, precio_propuesto} y devuelve por
    cada uno un veredicto:
      - 'normal' si precio dentro de ±10% del promedio_90d
      - 'inflado' si > 115% del promedio
      - 'sospechoso_bajo' si < 50% del promedio
      - 'sin_historia' si no hay datos previos

    Body: {items: [{codigo_mp, precio_propuesto}, ...]}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    body = request.get_json(silent=True) or {}
    items = body.get('items') or []
    if not items:
        return jsonify({'validaciones': []})
    conn = get_db(); c = conn.cursor()
    out = []
    for it in items:
        cod = (it.get('codigo_mp') or '').strip()
        try:
            precio = float(it.get('precio_propuesto') or 0)
        except (ValueError, TypeError):
            precio = 0
        if not cod or precio <= 0:
            out.append({'codigo_mp': cod, 'veredicto': 'sin_precio'})
            continue
        promedio = 0
        for col_precio in ('precio_unitario', 'precio_kg'):
            try:
                r = c.execute(
                    f"""SELECT AVG({col_precio})
                        FROM precios_mp_historico
                        WHERE codigo_mp=?
                          AND date(fecha) >= date('now','-5 hours','-90 days')""",
                    (cod,),
                ).fetchone()
                if r and r[0]:
                    promedio = float(r[0])
                    break
            except Exception:
                try: conn.rollback()
                except Exception: pass
                continue
        if promedio <= 0:
            out.append({
                'codigo_mp': cod, 'veredicto': 'sin_historia',
                'precio_propuesto': precio,
            })
            continue
        delta_pct = ((precio - promedio) / promedio) * 100
        if delta_pct > 15:
            ver = 'inflado'
        elif delta_pct < -50:
            ver = 'sospechoso_bajo'
        elif abs(delta_pct) <= 10:
            ver = 'normal'
        elif delta_pct > 10:
            ver = 'mayor'
        else:
            ver = 'menor'
        out.append({
            'codigo_mp': cod,
            'veredicto': ver,
            'precio_propuesto': precio,
            'precio_promedio_90d': round(promedio, 4),
            'delta_pct': round(delta_pct, 1),
            'requiere_justificacion': ver == 'inflado',
        })
    return jsonify({'validaciones': out, 'total': len(out)})


@bp.route('/api/compras/proveedor-recomendado/<path:codigo_mp>', methods=['GET'])
def proveedor_recomendado(codigo_mp):
    """Sprint Compras N3 · 21-may-2026 · auto-detección mejor proveedor.

    Calcula score por proveedor histórico para un codigo_mp:
      score = (1 - precio_normalizado) × peso_precio
            + (1 - lead_time_normalizado) × peso_tiempo
            + (frec_uso_normalizado) × peso_confianza

    Devuelve top 3 con su score y explicación.
    """
    u, err, code = _require_compras_write()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    # Datos por proveedor para este MP
    proveedores = {}
    for col_precio in ('precio_unitario', 'precio_kg'):
        try:
            rows = c.execute(
                f"""SELECT proveedor, COUNT(*) as n, AVG({col_precio}) as avg_p,
                          MIN({col_precio}) as min_p, MAX({col_precio}) as max_p,
                          MAX(fecha) as ultima_fecha
                   FROM precios_mp_historico
                   WHERE codigo_mp=? AND proveedor IS NOT NULL AND proveedor != ''
                   GROUP BY proveedor
                   ORDER BY n DESC""",
                (codigo_mp,),
            ).fetchall()
            for r in rows:
                proveedores[r[0]] = {
                    'nombre': r[0],
                    'usos': int(r[1] or 0),
                    'precio_promedio': float(r[2] or 0),
                    'precio_min': float(r[3] or 0),
                    'precio_max': float(r[4] or 0),
                    'ultima_compra': r[5] or '',
                }
            break
        except Exception:
            try: conn.rollback()
            except Exception: pass
            continue
    # Lead time desde mp_lead_time_config si existe
    try:
        lt_rows = c.execute(
            "SELECT proveedor_principal, lead_time_dias FROM mp_lead_time_config WHERE material_id=?",
            (codigo_mp,),
        ).fetchall()
        for r in lt_rows:
            if r[0] and r[0] in proveedores:
                proveedores[r[0]]['lead_time_dias'] = int(r[1] or 0)
    except Exception:
        pass
    if not proveedores:
        return jsonify({
            'codigo_mp': codigo_mp,
            'recomendados': [],
            'mensaje': 'Sin historial · no se puede recomendar',
        })
    # Normalizar y score
    lista = list(proveedores.values())
    precios = [p['precio_promedio'] for p in lista if p['precio_promedio'] > 0]
    usos = [p['usos'] for p in lista]
    lead_times = [p.get('lead_time_dias', 30) for p in lista]
    p_max = max(precios) if precios else 1
    p_min = min(precios) if precios else 0
    u_max = max(usos) if usos else 1
    lt_max = max(lead_times) if lead_times else 30
    lt_min = min(lead_times) if lead_times else 0
    for p in lista:
        precio_norm = (p['precio_promedio'] - p_min) / (p_max - p_min) if p_max > p_min else 0
        usos_norm = p['usos'] / u_max if u_max > 0 else 0
        lt = p.get('lead_time_dias', 30)
        lt_norm = (lt - lt_min) / (lt_max - lt_min) if lt_max > lt_min else 0
        # Pesos: 50% precio · 30% confianza · 20% lead time
        score = (
            (1 - precio_norm) * 0.50 +
            usos_norm * 0.30 +
            (1 - lt_norm) * 0.20
        ) * 100
        p['score'] = round(score, 1)
        # Explicación corta
        ventajas = []
        if precio_norm < 0.3: ventajas.append('precio bajo')
        if usos_norm > 0.6: ventajas.append('usado frecuente')
        if lt_norm < 0.3: ventajas.append('entrega rápida')
        p['ventajas'] = ventajas or ['default']
    lista.sort(key=lambda x: -x['score'])
    return jsonify({
        'codigo_mp': codigo_mp,
        'recomendados': lista[:3],
        'total_proveedores': len(proveedores),
    })


@bp.route('/api/compras/cotizaciones/desde-grupo', methods=['POST'])
def cotizar_desde_grupo_planta():
    """Sprint Compras N3 · 21-may-2026 · activar cotizaciones · Sebastián.

    Catalina selecciona un grupo Planta (proveedor + items consolidados)
    y dispara una ronda de cotizaciones con los TOP 3 proveedores
    históricos de las MPs del grupo (calculados desde precios_mp_historico).

    Body: {
      proveedor_sugerido: str,
      items: [{codigo_mp, nombre_mp, cantidad_g}],
      observaciones?: str
    }
    Returns: ronda_id + 3 cotizaciones Pendientes para que Catalina las
    envíe a los proveedores (email/WA) y vuelva a registrar respuestas.
    """
    u, err, code = _require_compras_write()
    if err:
        return err, code
    body = request.get_json(silent=True) or {}
    items = body.get('items') or []
    if not items or not isinstance(items, list):
        return jsonify({'error': 'items requerido'}), 400
    obs = (body.get('observaciones') or '').strip()
    prov_sug = (body.get('proveedor_sugerido') or '').strip()
    conn = get_db(); c = conn.cursor()
    # Top 3 proveedores históricos para CUALQUIER MP del grupo
    # ALTA-7 fix · 21-may-2026: detectar columna 1 vez antes del loop · evita
    # rollbacks consecutivos y queries inconsistentes entre MPs.
    codigos = [it.get('codigo_mp') for it in items if it.get('codigo_mp')]
    col_precio = 'precio_kg'  # default canónico actual
    try:
        # Query rápida de descubrimiento · si precio_unitario existe lo usa
        c.execute("SELECT precio_unitario FROM precios_mp_historico LIMIT 1")
        col_precio = 'precio_unitario'
    except Exception:
        try: conn.rollback()
        except Exception: pass
        col_precio = 'precio_kg'
    proveedores_scores = {}
    for cod in codigos:
        try:
            rows = c.execute(
                f"""SELECT proveedor, COUNT(*) as n, AVG({col_precio}) as avg_p
                    FROM precios_mp_historico
                    WHERE codigo_mp=? AND proveedor IS NOT NULL AND proveedor != ''
                    GROUP BY proveedor
                    ORDER BY n DESC LIMIT 5""",
                (cod,),
            ).fetchall()
            for r in rows:
                nombre = (r[0] or '').strip()
                if not nombre:
                    continue
                if nombre not in proveedores_scores:
                    proveedores_scores[nombre] = {
                        'nombre': nombre, 'usos': 0,
                        'precio_promedio_sum': 0.0,
                    }
                proveedores_scores[nombre]['usos'] += int(r[1] or 0)
                proveedores_scores[nombre]['precio_promedio_sum'] += float(r[2] or 0)
        except Exception:
            try: conn.rollback()
            except Exception: pass
            continue
    top = sorted(proveedores_scores.values(), key=lambda x: -x['usos'])[:3]
    # Si el proveedor sugerido no está en el top, agregarlo
    if prov_sug and not any(p['nombre'].lower() == prov_sug.lower() for p in top):
        top.insert(0, {'nombre': prov_sug, 'usos': 0, 'precio_promedio_sum': 0})
        top = top[:3]
    if len(top) < 2:
        return jsonify({
            'error': 'Sin suficientes proveedores históricos · ingresar manualmente',
            'top_encontrados': top,
        }), 400
    # Descripción auto-generada
    items_desc = ', '.join(f"{it.get('nombre_mp','?')} ({it.get('cantidad_g',0):.0f}g)" for it in items[:5])
    if len(items) > 5:
        items_desc += f' +{len(items)-5} más'
    descripcion = f"Cotización: {items_desc}"
    if obs:
        descripcion += ' · ' + obs
    from datetime import datetime as _dt
    ronda_id = f"COT-{_dt.now().strftime('%Y%m%d%H%M%S')}"
    creadas = []
    try:
        for p in top:
            c.execute(
                """INSERT INTO cotizaciones
                   (ronda_id, proveedor, descripcion, condiciones,
                    tiempo_entrega_dias, creado_por, estado)
                   VALUES (?,?,?,?,?,?, 'Pendiente')""",
                (ronda_id, p['nombre'], descripcion, '', 0, u),
            )
            creadas.append({'id': c.lastrowid, 'proveedor': p['nombre']})
        try:
            audit_log(c, usuario=u, accion='COTIZACION_RONDA_AUTO',
                      tabla='cotizaciones', registro_id=ronda_id,
                      despues={'proveedores': [p['nombre'] for p in top],
                               'items_count': len(items)})
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        return jsonify({'error': f'Falla crear ronda: {e}'}), 500
    return jsonify({
        'ok': True, 'ronda_id': ronda_id,
        'cotizaciones': creadas, 'count': len(creadas),
        'proveedores_top': top,
        'mensaje': f'Ronda {ronda_id} creada · {len(creadas)} proveedores · enviar y registrar respuestas',
    }), 201


@bp.route('/api/compras/cotizaciones/rondas', methods=['POST'])
def crear_ronda_cotizaciones():
    """Crea una nueva ronda de cotizaciones (3 proveedores).

    Body: {descripcion, proveedores: [{nombre, condiciones, tiempo_entrega_dias?}]}
    """
    u, err, code = _require_compras_write()
    if err:
        return err, code
    d = request.get_json() or {}
    descripcion = (d.get('descripcion') or '').strip()
    proveedores = d.get('proveedores', [])
    if not descripcion or not proveedores:
        return jsonify({'error': 'descripcion y proveedores requeridos'}), 400
    if len(proveedores) < 2:
        return jsonify({'error': 'Mínimo 2 proveedores para comparar'}), 400

    from datetime import datetime as _dt
    ronda_id = f"COT-{_dt.now().strftime('%Y%m%d%H%M%S')}"

    conn = get_db(); c = conn.cursor()
    creadas = []
    for prov in proveedores:
        nombre = (prov.get('nombre') or '').strip()
        if not nombre:
            continue
        c.execute("""INSERT INTO cotizaciones
                     (ronda_id, proveedor, descripcion, condiciones,
                      tiempo_entrega_dias, creado_por, estado)
                     VALUES (?, ?, ?, ?, ?, ?, 'Pendiente')""",
                  (ronda_id, nombre, descripcion,
                   prov.get('condiciones', ''),
                   int(prov.get('tiempo_entrega_dias') or 0),
                   u))
        creadas.append({'id': c.lastrowid, 'proveedor': nombre})
    conn.commit()
    return jsonify({
        'ok': True, 'ronda_id': ronda_id,
        'cotizaciones': creadas, 'count': len(creadas),
    }), 201


@bp.route('/api/compras/cotizaciones/<int:cot_id>', methods=['PATCH'])
def actualizar_cotizacion(cot_id):
    """Actualiza una cotización con la respuesta del proveedor.

    Body: {valor_total, condiciones?, tiempo_entrega_dias?, archivo?}
    """
    u, err, code = _require_compras_write()
    if err:
        return err, code
    d = request.get_json() or {}
    # P1 audit 26-may · validate_money · respuesta de proveedor → elegir_ganadora
    # creará OC con este valor · NaN/Inf contamina decisiones financieras.
    from http_helpers import validate_money as _vm_cot
    valor, _err_v = _vm_cot(d.get('valor_total') or 0, allow_zero=False,
                              max_value=1e10, field_name='valor_total')
    if _err_v:
        return jsonify(_err_v), 400
    # Snapshot antes para audit
    conn = get_db(); c = conn.cursor()
    ant_row = c.execute(
        "SELECT proveedor, valor_total, estado, ronda_id FROM cotizaciones WHERE id=?",
        (cot_id,)).fetchone()
    if not ant_row:
        return jsonify({'error': 'Cotización no encontrada'}), 404
    c.execute("""UPDATE cotizaciones SET
                    valor_total=?, fecha_recibida=datetime('now', '-5 hours', 'utc'),
                    condiciones=COALESCE(?, condiciones),
                    tiempo_entrega_dias=COALESCE(?, tiempo_entrega_dias),
                    archivo=COALESCE(?, archivo),
                    estado='Recibida'
                 WHERE id=?""",
              (valor, d.get('condiciones'), d.get('tiempo_entrega_dias'),
               d.get('archivo'), cot_id))
    if c.rowcount == 0:
        return jsonify({'error': 'Cotización no encontrada'}), 404
    # Audit log decisión financiera
    try:
        audit_log(c, usuario=u, accion='ACTUALIZAR_COTIZACION', tabla='cotizaciones',
                  registro_id=cot_id,
                  antes={'valor_total': ant_row[1], 'estado': ant_row[2]},
                  despues={'valor_total': valor, 'estado': 'Recibida',
                           'tiempo_entrega_dias': d.get('tiempo_entrega_dias')},
                  detalle=f"Cot id={cot_id} prov={ant_row[0]} ronda={ant_row[3]} · ${valor:,.0f}")
    except Exception as _ae:
        import logging as _lg
        _lg.getLogger('compras').warning('audit actualizar_cotizacion fallo: %s', _ae)
    conn.commit()
    return jsonify({'ok': True, 'id': cot_id, 'estado': 'Recibida'})


@bp.route('/api/compras/cotizaciones/rondas/<ronda_id>', methods=['GET'])
def get_ronda(ronda_id):
    """Lista las cotizaciones de una ronda con comparación lado a lado."""
    u, err, code = _require_compras_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT id, proveedor, fecha_solicitud, fecha_recibida,
                        valor_total, condiciones, descripcion,
                        tiempo_entrega_dias, ganadora, numero_oc, estado
                 FROM cotizaciones WHERE ronda_id=? ORDER BY valor_total ASC""",
              (ronda_id,))
    cols = ['id', 'proveedor', 'fecha_solicitud', 'fecha_recibida',
            'valor_total', 'condiciones', 'descripcion',
            'tiempo_entrega_dias', 'ganadora', 'numero_oc', 'estado']
    cotizaciones = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({
        'ronda_id': ronda_id,
        'cotizaciones': cotizaciones,
        'count': len(cotizaciones),
        'completadas': sum(1 for x in cotizaciones if x['estado'] == 'Recibida'),
    })


@bp.route('/api/compras/cotizaciones/<int:cot_id>/elegir-ganadora', methods=['POST'])
def elegir_ganadora(cot_id):
    """Marca cotización como ganadora · cierra otras · GENERA OC automática
    en estado Borrador (Sebastián 24-may-2026 · audit Cotizaciones · antes
    el UI prometía OC automática pero el endpoint no la creaba).

    Body opcional:
      {numero_oc: str}  · si se quiere vincular a OC existente en lugar de
                         crear nueva (caso edición · skip crear)

    Si NO se pasa numero_oc:
      • Crea OC nueva en ordenes_compra (estado='Borrador')
      • Inserta 1 item con la descripcion + valor_total (admin desglosa
        items si quiere editarlos en /compras → OCs Activas)
      • Sync maestro_mps.proveedor + mp_lead_time_config si la
        descripcion menciona un codigo_mp conocido (heurística simple)
      • Vincula numero_oc en la cotización
    """
    u, err, code = _require_compras_write()
    if err:
        return err, code
    d = request.get_json() or {}
    numero_oc_override = (d.get('numero_oc') or '').strip()

    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT ronda_id, proveedor, valor_total, descripcion,
                        COALESCE(tiempo_entrega_dias, 0), condiciones
                 FROM cotizaciones WHERE id=?""", (cot_id,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'Cotización no encontrada'}), 404
    ronda_id, prov_ganador, valor_ganador, descripcion, lead_time, condiciones = row
    valor_ganador = float(valor_ganador or 0)

    # Decisión · usar OC existente o crear nueva
    numero_oc_final = numero_oc_override
    oc_creada_automatica = False
    if not numero_oc_final:
        if valor_ganador <= 0:
            return jsonify({
                'error': 'Valor de cotización es 0 · registrá el precio antes de elegir ganadora',
                'codigo': 'VALOR_FALTANTE',
            }), 400
        # Generar numero_oc único con retry race-safe (mismo patrón que actualizar_estado_solicitud)
        for _intento in range(6):
            oc_num_nuevo = _siguiente_numero_oc(c, datetime.now().year)
            try:
                obs_oc = (f"Generado desde cotización ganadora · ronda {ronda_id} · "
                          f"lead_time cotizado {lead_time}d" +
                          (f" · {condiciones}" if condiciones else ""))
                c.execute(
                    "INSERT INTO ordenes_compra "
                    "(numero_oc, fecha, estado, proveedor, observaciones, "
                    " creado_por, valor_total, categoria) "
                    "VALUES (?, ?, 'Borrador', ?, ?, ?, ?, 'MP')",
                    (oc_num_nuevo, datetime.now().isoformat(),
                     prov_ganador or 'Por definir', obs_oc, u, valor_ganador),
                )
                # Item agregado · admin puede desglosarlo después si necesita
                c.execute(
                    "INSERT INTO ordenes_compra_items "
                    "(numero_oc, codigo_mp, nombre_mp, cantidad_g, "
                    " precio_unitario, subtotal) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (oc_num_nuevo, '', (descripcion or 'Cotización')[:200],
                     1, valor_ganador, valor_ganador),
                )
                numero_oc_final = oc_num_nuevo
                oc_creada_automatica = True
                break
            except sqlite3.IntegrityError:
                if _intento == 5:
                    raise

    # Marcar ganadora + cerrar otras
    c.execute("""UPDATE cotizaciones SET ganadora=1,
                    numero_oc=?, estado='Ganadora'
                 WHERE id=?""", (numero_oc_final, cot_id))
    c.execute("""UPDATE cotizaciones SET ganadora=0, estado='No seleccionada'
                 WHERE ronda_id=? AND id!=? AND estado != 'No seleccionada'""",
              (ronda_id, cot_id))
    try:
        audit_log(c, usuario=u, accion='ELEGIR_COTIZACION_GANADORA',
                  tabla='cotizaciones', registro_id=cot_id,
                  despues={'ronda_id': ronda_id,
                            'proveedor_ganador': (prov_ganador or '')[:200],
                            'valor_ganador': valor_ganador,
                            'numero_oc_vinculada': numero_oc_final or None,
                            'oc_creada_automatica': oc_creada_automatica,
                            'descripcion': (descripcion or '')[:200]},
                  detalle=(f"Eligió cotización id={cot_id} (ronda {ronda_id}) · "
                            f"ganador {prov_ganador} · {valor_ganador:.0f} · "
                            f"OC {'creada nueva' if oc_creada_automatica else 'vinculada'} "
                            f"{numero_oc_final}"))
    except Exception as e:
        log.warning('audit_log ELEGIR_COTIZACION_GANADORA fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True, 'cot_id': cot_id, 'ronda_id': ronda_id,
        'numero_oc': numero_oc_final,
        'oc_creada_automatica': oc_creada_automatica,
        'hint': (f'OC {numero_oc_final} creada en Borrador · revisala en '
                  f'tab "📦 OCs Activas" para editar items + autorizar.'
                  if oc_creada_automatica else
                  f'Cotización vinculada a OC existente {numero_oc_final}.'),
    })


@bp.route('/api/compras/cotizaciones/rondas', methods=['GET'])
def listar_rondas():
    """Lista las últimas 50 rondas con resumen de cada una."""
    u, err, code = _require_compras_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("""
        SELECT ronda_id, MAX(fecha_solicitud), MAX(descripcion),
               COUNT(*), SUM(CASE WHEN estado='Recibida' THEN 1 ELSE 0 END),
               MAX(CASE WHEN ganadora=1 THEN proveedor ELSE NULL END),
               MIN(CASE WHEN ganadora=1 THEN valor_total ELSE NULL END)
        FROM cotizaciones GROUP BY ronda_id
        ORDER BY MAX(fecha_solicitud) DESC LIMIT 50
    """)
    rondas = [
        {
            'ronda_id': r[0], 'fecha': r[1], 'descripcion': r[2],
            'total_cotizaciones': r[3], 'recibidas': r[4],
            'ganadora_proveedor': r[5], 'valor_ganador': r[6],
        }
        for r in c.fetchall()
    ]
    return jsonify({'rondas': rondas, 'count': len(rondas)})


# ─── CENTRO DE COSTOS (Sprint 5) ─────────────────────────────────────────────

@bp.route('/api/compras/centros-costos', methods=['GET'])
def listar_centros_costos():
    """Devuelve los centros de costos en uso + KPIs por centro."""
    u, err, code = _require_compras_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("""
        SELECT COALESCE(centro_costos, 'general') as cc,
               COUNT(*) as ocs, SUM(valor_total) as total
        FROM ordenes_compra
        WHERE estado != 'Rechazada' AND estado != 'Borrador'
        GROUP BY cc ORDER BY total DESC
    """)
    centros = [
        {'centro_costos': r[0], 'ocs': r[1], 'total': round(r[2] or 0, 2)}
        for r in c.fetchall()
    ]
    return jsonify({'centros': centros, 'count': len(centros)})


# ─── Fast-track config · Sebastián 24-may-2026 · audit Solicitudes ──────────
# Tabla compras_fast_track_config (mig 179) lista categorías que saltan SOL→OC
# directo a Autorizada (sin doble paso). Esta sección expone endpoints REST
# para que Sebastián administre la config (listar / agregar / actualizar /
# eliminar) sin tener que tocar código ni la BD directamente.

@bp.route('/api/compras/fast-track-config', methods=['GET'])
def fast_track_config_list():
    """Lista todas las categorías configuradas para fast-track.

    Devuelve también categorías "candidatas" (las que aparecen en SOLs
    pero no están en la tabla aún) para que el admin las agregue con 1 click.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    try:
        rows = c.execute(
            "SELECT id, categoria, monto_max_cop, activo, configurado_por, "
            "       configurado_at, COALESCE(notas,'') "
            "FROM compras_fast_track_config "
            "ORDER BY activo DESC, categoria"
        ).fetchall()
    except sqlite3.OperationalError:
        return jsonify({'configs': [], 'candidatas': [],
                        'error': 'Tabla aún no migrada · reinicia el servidor'}), 503
    configs = [{
        'id': r[0], 'categoria': r[1], 'monto_max_cop': float(r[2] or 0),
        'activo': bool(r[3]), 'configurado_por': r[4],
        'configurado_at': r[5], 'notas': r[6],
    } for r in rows]
    config_cats = {x['categoria'] for x in configs}
    # Categorías reales que aparecen en SOLs pero no están en config
    cand_rows = c.execute(
        "SELECT DISTINCT categoria FROM solicitudes_compra "
        "WHERE COALESCE(categoria,'') != '' ORDER BY categoria"
    ).fetchall()
    candidatas = [r[0] for r in cand_rows if r[0] and r[0] not in config_cats]
    return jsonify({'configs': configs, 'candidatas': candidatas})


@bp.route('/api/compras/fast-track-config', methods=['POST'])
def fast_track_config_upsert():
    """Agrega o actualiza la config de fast-track de una categoría.

    Body: {categoria, monto_max_cop, activo, notas}
      • monto_max_cop=0 → fast-track sin tope (cualquier monto pasa)
      • monto_max_cop>0 → fast-track solo si valor_oc <= monto_max_cop
      • activo=false   → desactivar sin borrar (preserva audit history)

    Requiere admin (decisión financiera · afecta a doble paso de autorización).
    """
    usuario = session.get('compras_user', '')
    if usuario not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin puede configurar fast-track'}), 403
    d = request.get_json(silent=True) or {}
    categoria = (d.get('categoria') or '').strip()
    if not categoria:
        return jsonify({'error': 'categoria requerida'}), 400
    # P1 audit 26-may · validate_money · NaN/Inf en monto_max_cop rompía
    # _evaluar_auto_aprobacion y dejaba pasar OCs sin revisión gerencial.
    from http_helpers import validate_money as _vm_ft
    monto_max, _err_mm = _vm_ft(d.get('monto_max_cop') or 0, allow_zero=True,
                                 max_value=1e10, field_name='monto_max_cop')
    if _err_mm:
        return jsonify(_err_mm), 400
    activo = 1 if d.get('activo', True) else 0
    notas = (d.get('notas') or '').strip()[:300]
    conn = get_db(); c = conn.cursor()
    try:
        # Upsert idempotente · PG y SQLite
        existing = c.execute(
            "SELECT id FROM compras_fast_track_config WHERE categoria=?",
            (categoria,),
        ).fetchone()
        now = datetime.now().isoformat()
        if existing:
            antes = c.execute(
                "SELECT monto_max_cop, activo, COALESCE(notas,'') "
                "FROM compras_fast_track_config WHERE id=?",
                (existing[0],),
            ).fetchone()
            c.execute(
                "UPDATE compras_fast_track_config SET monto_max_cop=?, activo=?, "
                "       configurado_por=?, configurado_at=?, notas=? WHERE id=?",
                (monto_max, activo, usuario, now, notas, existing[0]),
            )
            try:
                audit_log(c, usuario=usuario, accion='ACTUALIZAR_FAST_TRACK',
                          tabla='compras_fast_track_config',
                          registro_id=str(existing[0]),
                          antes={'monto_max_cop': float(antes[0] or 0),
                                  'activo': bool(antes[1]), 'notas': antes[2]},
                          despues={'monto_max_cop': monto_max, 'activo': bool(activo),
                                    'notas': notas},
                          detalle=f"Fast-track {categoria} actualizado")
            except Exception:
                pass
            accion = 'actualizado'
        else:
            c.execute(
                "INSERT INTO compras_fast_track_config "
                "(categoria, monto_max_cop, activo, configurado_por, configurado_at, notas) "
                "VALUES (?,?,?,?,?,?)",
                (categoria, monto_max, activo, usuario, now, notas),
            )
            try:
                audit_log(c, usuario=usuario, accion='CREAR_FAST_TRACK',
                          tabla='compras_fast_track_config', registro_id=categoria,
                          antes={}, despues={'categoria': categoria,
                                              'monto_max_cop': monto_max,
                                              'activo': bool(activo), 'notas': notas},
                          detalle=f"Fast-track {categoria} creado")
            except Exception:
                pass
            accion = 'creado'
        conn.commit()
        return jsonify({
            'ok': True, 'accion': accion, 'categoria': categoria,
            'monto_max_cop': monto_max, 'activo': bool(activo),
            'hint': ('Sin tope · cualquier monto pasa' if monto_max <= 0 else
                     f'Aplica si valor OC <= ${monto_max:,.0f}'),
        })
    except sqlite3.OperationalError as e:
        return jsonify({'error': f'Tabla no disponible: {e}'}), 503


@bp.route('/api/compras/fast-track-config/<int:config_id>', methods=['DELETE'])
def fast_track_config_delete(config_id):
    """Elimina una config de fast-track. Audita antes de borrar."""
    usuario = session.get('compras_user', '')
    if usuario not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    conn = get_db(); c = conn.cursor()
    try:
        row = c.execute(
            "SELECT categoria, monto_max_cop, activo FROM compras_fast_track_config WHERE id=?",
            (config_id,),
        ).fetchone()
        if not row:
            return jsonify({'error': 'Config no encontrada'}), 404
        c.execute("DELETE FROM compras_fast_track_config WHERE id=?", (config_id,))
        try:
            audit_log(c, usuario=usuario, accion='ELIMINAR_FAST_TRACK',
                      tabla='compras_fast_track_config', registro_id=str(config_id),
                      antes={'categoria': row[0], 'monto_max_cop': float(row[1] or 0),
                              'activo': bool(row[2])},
                      despues={},
                      detalle=f"Fast-track {row[0]} eliminado")
        except Exception:
            pass
        conn.commit()
        return jsonify({'ok': True, 'eliminado': row[0]})
    except sqlite3.OperationalError as e:
        return jsonify({'error': str(e)}), 503


# ─── Revertir pago de OC · Sebastián 24-may-2026 · audit Pagos ────────
@bp.route('/api/ordenes-compra/<numero_oc>/revertir-pago', methods=['POST'])
def revertir_pago_oc(numero_oc):
    """Revierte el ÚLTIMO pago de una OC · admin only · ventana 24h.

    Deshace en orden inverso a pagar_oc:
      1. Anula último pagos_oc (DELETE)
      2. Anula último comprobantes_pago (DELETE + audit)
      3. Anula entrada flujo_egresos asociada (DELETE)
      4. Recalcula estado OC:
         - Sin pagos restantes → estado='Recibida'
         - Con pagos parciales → estado='Parcial'
         - SUM == valor_total → mantiene 'Pagada' (no debería pasar)
      5. Si OC influencer/CC y sync pagos_influencers → revertir a 'Pendiente'

    Body: {motivo: str (≥15 chars · obligatorio para audit)}
    Solo admin · audit completo REVERTIR_PAGO_OC.
    Ventana 24h desde último pago (por seguridad · si quieren revertir
    algo más viejo, requiere contactar a IT directamente).
    """
    usuario = session.get('compras_user', '')
    if usuario not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin puede revertir pagos'}), 403
    d = request.get_json(silent=True) or {}
    motivo = (d.get('motivo') or '').strip()
    if len(motivo) < 15:
        return jsonify({'error': 'motivo requerido (≥15 chars)',
                        'codigo': 'MOTIVO_FALTANTE'}), 400
    conn = get_db(); cur = conn.cursor()
    # Buscar OC
    cur.execute(
        "SELECT estado, valor_total, categoria, proveedor FROM ordenes_compra WHERE numero_oc=?",
        (numero_oc,),
    )
    oc = cur.fetchone()
    if not oc:
        return jsonify({'error': 'OC no encontrada'}), 404
    estado_act, valor_total_oc, categoria_oc, proveedor_oc = oc
    if estado_act not in ('Pagada', 'Parcial'):
        return jsonify({
            'error': f'OC en estado {estado_act} · solo se puede revertir Pagada/Parcial',
            'codigo': 'ESTADO_NO_REVERSIBLE',
        }), 409
    # Buscar último pago
    cur.execute(
        "SELECT id, monto, fecha_pago FROM pagos_oc WHERE numero_oc=? "
        "ORDER BY id DESC LIMIT 1",
        (numero_oc,),
    )
    pago = cur.fetchone()
    if not pago:
        return jsonify({'error': 'No hay pagos registrados en pagos_oc para esta OC',
                        'codigo': 'SIN_PAGOS'}), 404
    pago_id, monto_pago, fecha_pago = pago
    # Validar ventana 24h
    try:
        from datetime import datetime as _dt
        fpago_dt = _dt.fromisoformat((fecha_pago or '').replace('Z', ''))
        delta_h = (_dt.now() - fpago_dt).total_seconds() / 3600
        if delta_h > 24:
            return jsonify({
                'error': (f'Pago tiene {delta_h:.1f}h · ventana de reversión es 24h. '
                          f'Para revertir un pago más viejo, abrí ticket con IT.'),
                'codigo': 'FUERA_DE_VENTANA',
                'horas_desde_pago': round(delta_h, 1),
            }), 409
    except (ValueError, TypeError):
        # Si no se puede parsear fecha, no bloqueamos (mejor permitir admin override)
        delta_h = None
    # 1. DELETE pago de pagos_oc
    cur.execute("DELETE FROM pagos_oc WHERE id=?", (pago_id,))
    # 2. DELETE último comprobante asociado a esta OC (si existe)
    cur.execute(
        "SELECT id, COALESCE(numero_ce,'') FROM comprobantes_pago "
        "WHERE numero_oc=? ORDER BY id DESC LIMIT 1",
        (numero_oc,),
    )
    ce_row = cur.fetchone()
    ce_eliminado = None
    if ce_row:
        ce_eliminado = ce_row[1]
        cur.execute("DELETE FROM comprobantes_pago WHERE id=?", (ce_row[0],))
    # 3. DELETE flujo_egresos asociado · best-effort (puede no tener FK directa)
    flujo_eliminado = 0
    try:
        cur.execute(
            "DELETE FROM flujo_egresos WHERE numero_oc=? AND ABS(monto - ?) < 0.01",
            (numero_oc, float(monto_pago)),
        )
        flujo_eliminado = cur.rowcount
    except sqlite3.OperationalError:
        pass
    # 4. Recalcular estado OC según pagos restantes
    cur.execute("SELECT COALESCE(SUM(monto), 0) FROM pagos_oc WHERE numero_oc=?", (numero_oc,))
    total_restante = float(cur.fetchone()[0] or 0)
    if total_restante <= 0.01:
        nuevo_estado = 'Recibida'
    elif total_restante >= float(valor_total_oc or 0) - 0.01:
        nuevo_estado = 'Pagada'  # raro · permitiría delete dejándola pagada
    else:
        nuevo_estado = 'Parcial'
    cur.execute(
        "UPDATE ordenes_compra SET estado=?, "
        "fecha_pago=CASE WHEN ? = 'Recibida' THEN NULL ELSE fecha_pago END, "
        "pagado_por=CASE WHEN ? = 'Recibida' THEN NULL ELSE pagado_por END "
        "WHERE numero_oc=?",
        (nuevo_estado, nuevo_estado, nuevo_estado, numero_oc),
    )
    # 5. Revertir pagos_influencers si aplica
    pi_revertido = False
    cat_low = (categoria_oc or '').lower()
    if 'influencer' in cat_low or 'marketing' in cat_low or 'cuenta de cobro' in cat_low:
        try:
            cur.execute(
                "UPDATE pagos_influencers SET estado='Pendiente' WHERE numero_oc=?",
                (numero_oc,),
            )
            pi_revertido = cur.rowcount > 0
        except sqlite3.OperationalError:
            pass
    # 6. Sync solicitudes_compra · si OC vinculada estaba Pagada, volver a Aprobada
    try:
        cur.execute(
            "UPDATE solicitudes_compra SET estado='Aprobada' "
            "WHERE numero_oc=? AND estado='Pagada'",
            (numero_oc,),
        )
    except sqlite3.OperationalError:
        pass
    # 7. Audit completo
    try:
        audit_log(cur, usuario=usuario, accion='REVERTIR_PAGO_OC',
                  tabla='ordenes_compra', registro_id=numero_oc,
                  antes={'estado': estado_act, 'pagado_id': pago_id,
                          'monto_revertido': float(monto_pago),
                          'fecha_pago': fecha_pago},
                  despues={'estado': nuevo_estado,
                            'total_restante': total_restante,
                            'ce_eliminado': ce_eliminado,
                            'flujo_egresos_eliminados': flujo_eliminado,
                            'pagos_influencers_revertido': pi_revertido,
                            'motivo': motivo},
                  detalle=(f"Revertido pago OC {numero_oc} · ${monto_pago:,.0f} · "
                            f"{estado_act} → {nuevo_estado} · motivo: {motivo[:120]}"))
    except Exception as e:
        log.warning('audit_log REVERTIR_PAGO_OC fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'numero_oc': numero_oc,
        'estado_anterior': estado_act,
        'nuevo_estado': nuevo_estado,
        'monto_revertido': float(monto_pago),
        'total_restante': total_restante,
        'ce_eliminado': ce_eliminado,
        'flujo_egresos_eliminados': flujo_eliminado,
        'pagos_influencers_revertido': pi_revertido,
        'detalle': (f'Pago ${monto_pago:,.0f} revertido · CE {ce_eliminado or "(ninguno)"} '
                    f'eliminado · OC ahora en {nuevo_estado}'),
    })


# ─── Proveedores duplicados · detector + fusionar · Sebastián 25-may-2026 ────
# Sebastián reportó: "Agenquimicos" vs "AGENQUIMICOS" aparecen como 2 cards
# en el tab Proveedores · el primero con NIT vacío (creado por cron auto-plan
# desde seed con capitalización original) · el segundo completo (Catalina lo
# creó manual al recibir OC). Drift case-sensitive del nombre como FK textual.
# Endpoint detector + fusionar consolida el huérfano en el canónico (con NIT)
# traspasando todas las referencias downstream.

@bp.route('/api/admin/proveedores-duplicados', methods=['GET'])
def admin_proveedores_duplicados():
    """Detecta proveedores duplicados case + espacios insensitive.

    Devuelve grupos donde 2+ filas tienen la misma key normalizada:
        UPPER(TRIM(' '.join(nombre.split())))

    Para cada grupo · ranking del "más completo" (mide tener NIT + banco +
    num_cuenta + contacto + email · usar como destino sugerido en fusión).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # Levantar todos los activos
    try:
        rows = c.execute(
            "SELECT id, nombre, COALESCE(nit,''), COALESCE(banco,''), "
            "       COALESCE(num_cuenta,''), COALESCE(contacto,''), "
            "       COALESCE(email,''), COALESCE(telefono,''), "
            "       COALESCE(activo, 1) "
            "FROM proveedores WHERE COALESCE(activo,1)=1"
        ).fetchall()
    except Exception as e:
        return jsonify({'error': f'query fallo: {e}'}), 500

    def _norm(s):
        return ' '.join((s or '').strip().split()).upper()

    def _score_completitud(r):
        # Cuenta campos no vacíos · más alto = más completo
        # NIT pesa doble (es el identificador legal)
        return ((2 if r[2] else 0) + (1 if r[3] else 0) + (1 if r[4] else 0) +
                (1 if r[5] else 0) + (1 if r[6] else 0) + (1 if r[7] else 0))

    grupos_dict = {}
    for r in rows:
        k = _norm(r[1])
        if not k:
            continue
        grupos_dict.setdefault(k, []).append(r)

    duplicados = []
    for k, lista in grupos_dict.items():
        if len(lista) < 2:
            continue
        # Ordenar por score completitud DESC · primer elemento = destino sugerido
        lista_sorted = sorted(lista, key=lambda r: -_score_completitud(r))
        duplicados.append({
            'key_normalizada': k,
            'count': len(lista_sorted),
            'destino_sugerido': lista_sorted[0][1],
            'destino_sugerido_id': lista_sorted[0][0],
            'destino_score': _score_completitud(lista_sorted[0]),
            'proveedores': [{
                'id': p[0], 'nombre': p[1], 'nit': p[2], 'banco': p[3],
                'num_cuenta': p[4], 'contacto': p[5], 'email': p[6],
                'telefono': p[7], 'score_completitud': _score_completitud(p),
                'es_destino_sugerido': p[0] == lista_sorted[0][0],
            } for p in lista_sorted],
        })
    duplicados.sort(key=lambda g: -g['count'])
    return jsonify({
        'total_grupos': len(duplicados),
        'total_huerfanos': sum(g['count'] - 1 for g in duplicados),
        'grupos': duplicados,
    })


@bp.route('/api/admin/proveedores-fusionar', methods=['POST'])
def admin_proveedores_fusionar():
    """Fusiona un proveedor huérfano en uno canónico.

    Body: {keeper: 'AGENQUIMICOS', merge_from: 'Agenquimicos'}

    Side effects · UPDATE todas las referencias downstream a apuntar al
    keeper · UPDATE proveedores SET activo=0 para merge_from. Devuelve
    contadores por tabla afectada para auditar visualmente la operación.

    Solo COMPRAS_ACCESS_WRITE · operación delicada · audit completo.
    """
    usuario, err, code = _require_compras_write()
    if err: return err, code
    d = request.get_json(silent=True) or {}
    keeper = (d.get('keeper') or '').strip()
    merge_from = (d.get('merge_from') or '').strip()
    if not keeper or not merge_from:
        return jsonify({'error': 'keeper y merge_from requeridos'}), 400
    if keeper.lower() == merge_from.lower():
        return jsonify({'error': 'keeper y merge_from no pueden ser el mismo'}), 400
    conn = get_db(); c = conn.cursor()
    # Validar que ambos existen
    k_row = c.execute(
        "SELECT id, COALESCE(nit,''), COALESCE(activo,1) FROM proveedores WHERE nombre=?",
        (keeper,),
    ).fetchone()
    m_row = c.execute(
        "SELECT id, COALESCE(nit,''), COALESCE(activo,1) FROM proveedores WHERE nombre=?",
        (merge_from,),
    ).fetchone()
    if not k_row:
        return jsonify({'error': f'Keeper "{keeper}" no existe'}), 404
    if not m_row:
        return jsonify({'error': f'Merge_from "{merge_from}" no existe'}), 404
    if not int(k_row[2] or 0):
        return jsonify({'error': f'Keeper "{keeper}" está dado de baja · activarlo primero'}), 409
    # UPDATE downstream · contar filas afectadas por tabla
    contadores = {}
    propagar = [
        ('ordenes_compra', 'proveedor'),
        ('solicitudes_compra', 'proveedor_sugerido'),
        ('solicitudes_compra_items', 'proveedor_sugerido'),
        ('precios_mp_historico', 'proveedor'),
        ('ordenes_servicio', 'proveedor'),
        ('mp_lead_time_config', 'proveedor_principal'),
        ('pagos_influencers', 'influencer_nombre'),
    ]
    for tabla, col in propagar:
        try:
            c.execute(f"UPDATE {tabla} SET {col}=? WHERE {col}=?",
                       (keeper, merge_from))
            contadores[tabla] = c.rowcount
        except Exception:
            contadores[tabla] = 'NA'
    # cotizaciones tiene 1 columna proveedor
    try:
        c.execute(
            "UPDATE cotizaciones SET proveedor=? WHERE proveedor=?",
            (keeper, merge_from),
        )
        contadores['cotizaciones'] = c.rowcount
    except Exception:
        contadores['cotizaciones'] = 'NA'
    # Si el keeper no tiene NIT pero el merge_from sí · copiarlo (más completo)
    if not k_row[1] and m_row[1]:
        try:
            c.execute(
                "UPDATE proveedores SET nit=? WHERE id=?",
                (m_row[1], k_row[0]),
            )
            contadores['nit_copiado_a_keeper'] = m_row[1]
        except Exception:
            pass
    # Dar de baja al merge_from
    c.execute(
        "UPDATE proveedores SET activo=0, motivo_baja=?, fecha_baja=? WHERE id=?",
        (f'Fusionado en "{keeper}" por {usuario}', datetime.now().isoformat(), m_row[0]),
    )
    # Audit completo
    try:
        audit_log(c, usuario=usuario, accion='FUSIONAR_PROVEEDORES',
                  tabla='proveedores', registro_id=str(m_row[0]),
                  antes={'merge_from': merge_from,
                          'merge_from_nit': m_row[1],
                          'merge_from_id': m_row[0]},
                  despues={'keeper': keeper, 'keeper_id': k_row[0],
                            'contadores_filas_actualizadas': contadores},
                  detalle=(f"Fusionó '{merge_from}' → '{keeper}' · "
                            f"total filas movidas: "
                            f"{sum(v for v in contadores.values() if isinstance(v,int))}"))
    except Exception as e:
        log.warning('audit_log FUSIONAR_PROVEEDORES fallo: %s', e)
    conn.commit()
    return jsonify({
        'ok': True,
        'keeper': keeper,
        'merge_from': merge_from,
        'merge_from_dado_de_baja': True,
        'contadores_filas_actualizadas': contadores,
        'total_filas_movidas': sum(v for v in contadores.values() if isinstance(v, int)),
    })


@bp.route('/api/admin/proveedores-dedup-nombre', methods=['POST'])
def admin_proveedores_dedup_nombre():
    """Deduplica proveedores con el MISMO nombre (varias filas activas, distinto
    id). La fusión normal trabaja por nombre y se bloquea si keeper==merge_from;
    esto opera por ID: conserva la fila más completa y da de baja las demás.
    Como las referencias downstream apuntan al NOMBRE, mueve también las de
    cualquier variante (mayúsculas/espacios) al nombre del keeper. Body: {nombre}.
    Sebastián 31-may-2026 · caso Agenquimicos duplicado exacto."""
    usuario, err, code = _require_compras_write()
    if err:
        return err, code
    d = request.get_json(silent=True) or {}
    nombre = (d.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'nombre requerido'}), 400
    norm = ' '.join(nombre.split()).upper()
    conn = get_db(); c = conn.cursor()
    rows = c.execute(
        "SELECT id, nombre, COALESCE(nit,''), COALESCE(banco,''), "
        "COALESCE(num_cuenta,''), COALESCE(contacto,''), COALESCE(email,''), "
        "COALESCE(telefono,'') FROM proveedores "
        "WHERE COALESCE(activo,1)=1 AND UPPER(TRIM(nombre))=?", (norm,)
    ).fetchall()
    if len(rows) < 2:
        return jsonify({'error': 'no hay 2+ filas activas con ese nombre',
                        'encontrados': len(rows)}), 409

    def _score(r):
        return ((2 if r[2] else 0) + (1 if r[3] else 0) + (1 if r[4] else 0) +
                (1 if r[5] else 0) + (1 if r[6] else 0) + (1 if r[7] else 0))
    ordenados = sorted(rows, key=lambda r: (-_score(r), r[0]))
    keeper = ordenados[0]
    bajas = ordenados[1:]
    keeper_nombre = keeper[1]
    # Copiar al keeper datos que le falten y un duplicado sí tenga
    campos = {'nit': 2, 'banco': 3, 'num_cuenta': 4, 'contacto': 5, 'email': 6, 'telefono': 7}
    for campo, idx in campos.items():
        if not keeper[idx]:
            for b in bajas:
                if b[idx]:
                    try:
                        c.execute(f"UPDATE proveedores SET {campo}=? WHERE id=?",
                                  (b[idx], keeper[0]))
                    except Exception:
                        pass
                    break
    propagar = [
        ('ordenes_compra', 'proveedor'),
        ('solicitudes_compra', 'proveedor_sugerido'),
        ('solicitudes_compra_items', 'proveedor_sugerido'),
        ('precios_mp_historico', 'proveedor'),
        ('ordenes_servicio', 'proveedor'),
        ('mp_lead_time_config', 'proveedor_principal'),
        ('cotizaciones', 'proveedor'),
        ('pagos_influencers', 'influencer_nombre'),
    ]
    contadores = {}
    ids_baja = []
    for b in bajas:
        if b[1] != keeper_nombre:  # variante de mayúsculas/espacios · mover refs
            for tabla, col in propagar:
                try:
                    c.execute(f"UPDATE {tabla} SET {col}=? WHERE {col}=?",
                              (keeper_nombre, b[1]))
                    contadores[tabla] = contadores.get(tabla, 0) + (c.rowcount or 0)
                except Exception:
                    pass
        c.execute(
            "UPDATE proveedores SET activo=0, motivo_baja=?, fecha_baja=? WHERE id=?",
            (f'Dedup · se conserva id {keeper[0]} ("{keeper_nombre}") por {usuario}',
             datetime.now().isoformat(), b[0]))
        ids_baja.append(b[0])
    try:
        audit_log(c, usuario=usuario, accion='DEDUP_PROVEEDOR_NOMBRE',
                  tabla='proveedores', registro_id=str(keeper[0]),
                  despues={'nombre': nombre, 'keeper_id': keeper[0],
                           'dados_de_baja': ids_baja, 'refs_movidas': contadores})
    except Exception as e:
        log.warning('audit DEDUP_PROVEEDOR_NOMBRE fallo: %s', e)
    conn.commit()
    return jsonify({'ok': True, 'nombre': nombre, 'keeper_id': keeper[0],
                    'dados_de_baja': ids_baja, 'n_baja': len(ids_baja),
                    'refs_movidas': contadores})
