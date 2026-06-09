"""
Blueprint: Centro de Programación
Endpoints:
  GET  /api/programacion/resumen        — dashboard principal (velocidad + stock + alertas + IA)
  GET  /api/programacion/velocidad      — velocidad Shopify por SKU (últimos 60 días)
  GET  /api/programacion/calendario     — eventos Google Calendar próximos 90 días
  GET  /api/programacion/stock-60d      — proyección consolidada 60 días por producto
  GET  /api/programacion/alertas        — alertas inteligentes (schedule + stock)
  POST /api/programacion/seed-formulas  — (dev) fuerza re-seed de fórmulas desde DB
  GET  /api/programacion/productos      — lista de productos con fórmulas

Dependencias de env:
  GOOGLE_API_KEY   — API key con acceso a Calendar API (calendar público o compartido)
  CALENDAR_ID      — Google Calendar ID de producción
  ANTHROPIC_API_KEY — para narrativa IA
"""

from flask import Blueprint, g, jsonify, request, session
from database import db_connect
import os, json, logging, sqlite3, urllib.request, urllib.error, urllib.parse
from datetime import datetime, timedelta, date, timezone
from database import get_db
from config import ADMIN_USERS, APP_BASE_URL, CALIDAD_USERS, COMPRAS_USERS, PLANTA_USERS, EBR_MODE
from inventario_helpers import stock_mp_total, stock_mp_disponible
from audit_helpers import audit_log

# Bug latente preexistente fixed 2-may-2026: el blueprint usaba `log.warning()`
# y `logger.warning()` sin tener el logger definido. Cualquier call a un branch
# de error fallaba con NameError. Audit zero-error.
log = logging.getLogger('programacion')
logger = log  # alias compatible con call-sites históricos

bp = Blueprint('programacion', __name__)


def _caller_puede_operar_produccion(c, user, evento_id, rol_requerido='cualquiera'):
    """Sebastián 19-may-2026 · BUG-1 audit Planta PERFECTA · cierra hueco
    crítico: hasta hoy /iniciar y /completar solo chequeaban login. Un
    operario podía POST con id de producción ajena y descontar MPs en
    nombre de otro.

    Retorna (ok: bool, error_response: tuple|None). Si ok=False, retornar
    error_response directamente desde el endpoint.

    Args:
      rol_requerido: 'cualquiera' (default · operario asignado a alguno
        de los 4 roles), 'dispensacion' (estricto · solo dispensador o
        admin/jefe puede), 'finalizar' (dispensación o jefe).

    Reglas:
      - user en ADMIN_USERS → siempre ok (Sebastián / Alejandro)
      - user es operarios_planta.es_jefe_produccion=1 → ok (Luis Enrique)
      - user mapea a un operario_id que cumple `rol_requerido`
      - en cualquier otro caso → 403

    P0-2/P0-3 23-may-PM · auditoría agente Kanban cazó:
      · LIKE 'u%' fallback permitía suplantar operario por prefix match
        (e.g. user='m' matcheaba Mayerlin/Maria/Manuel). ELIMINADO.
      · Cualquier operario asignado podía descontar MP. Ahora rol_requerido
        ='dispensacion' restringe /iniciar (descuento MP) solo al dispensador.
    """
    if not user:
        return False, (jsonify({'error': 'No autorizado'}), 401)
    if user in ADMIN_USERS:
        return True, None
    u = user.lower().strip()
    # Match exacto por nombre
    op_row = c.execute(
        "SELECT id, COALESCE(es_jefe_produccion,0) FROM operarios_planta "
        "WHERE LOWER(nombre) = ? AND COALESCE(activo,1) = 1 LIMIT 1",
        (u,),
    ).fetchone()
    if not op_row and len(u) >= 4:
        # match 2: primer-letra + apellido (smurillo → Sebastian Murillo)
        op_row = c.execute(
            """SELECT id, COALESCE(es_jefe_produccion,0) FROM operarios_planta
               WHERE LOWER(apellido) = ?
                 AND LOWER(SUBSTR(nombre, 1, 1)) = ?
                 AND COALESCE(activo,1) = 1 LIMIT 1""",
            (u[1:], u[0]),
        ).fetchone()
    # P0-3 23-may-PM · ELIMINADO el fallback `LIKE u%` que permitía
    # suplantar operarios (user='m' matcheaba al primer operario que
    # empezara con m alfabéticamente · brecha de seguridad).
    if not op_row:
        return False, (jsonify({
            'error': 'Tu usuario no está mapeado a un operario en planta. '
                     'Pídele acceso a Sebastián.',
        }), 403)
    op_id, es_jefe = int(op_row[0]), bool(op_row[1])
    if es_jefe:
        return True, None
    rol_row = c.execute(
        """SELECT operario_dispensacion_id, operario_elaboracion_id,
                  operario_envasado_id, operario_acondicionamiento_id
           FROM produccion_programada WHERE id = ?""",
        (evento_id,),
    ).fetchone()
    if not rol_row:
        return False, (jsonify({'error': 'produccion no existe'}), 404)
    # P0-2 · validar rol específico cuando se requiere
    if rol_requerido == 'dispensacion':
        if op_id == rol_row[0]:
            return True, None
        return False, (jsonify({
            'error': 'Solo el operario de dispensación (o jefe/admin) puede '
                     'iniciar la producción y descontar MP. Tu rol asignado '
                     'es distinto.',
            'codigo': 'rol_incorrecto',
        }), 403)
    if op_id in [r for r in rol_row if r is not None]:
        return True, None
    return False, (jsonify({
        'error': 'No estás asignado a esta producción · pídele al admin '
                 'que te asigne o se la pase a otro operario.',
        'codigo': 'no_asignado',
    }), 403)

# ─────────────────────────────────────────────────────────────────────────────
# Sebastián 12-may-2026: constante global NON_FAB_KW unificada.
# Antes había 4 definiciones duplicadas (líneas 938, 3937, 7103, 8414) con
# variantes mínimas. Ahora una sola fuente de verdad para que actualizaciones
# del filtro apliquen a TODO el módulo. Eventos cuyo título matchee uno de
# estos substrings (lowercase) NO se importan a produccion_programada porque
# no consumen MPs crudas (envasado/QC/etc) o no son eventos productivos
# (reuniones/aniversarios/etc).
#
# Si Calendar tiene eventos legítimos que están siendo rechazados por error,
# evaluar si el título tiene palabras genéricas que coinciden con esta lista.
# Ejemplo: 'BLOQUE DIA 1' rechazado correctamente; 'GLOSS Día' similar.
NON_FAB_KW_GLOBAL = frozenset({
    # Operaciones post-fabricación (no consumen MPs crudas)
    'envasado', 'acondicionamiento', 'micro qc',
    'control de calidad', 'dispensado', 'etiquetado', 'llenado',
    # Eventos no-producción
    'reunion', 'reunión', 'ajuste contractual', 'ajuste cor',
    'meeting', 'agenda', 'capacitaci', 'training',
    # Sebastián 12-may-2026: eventos administrativos del Calendar
    'aniversario', 'cumpleaño', 'vacacion', 'feriado', 'festivo',
    'inventario fisico', 'inventario físico', 'auditoria', 'auditoría',
    'orden del dia', 'orden del día', 'planeacion', 'planeación',
    'comite', 'comité', 'visita', 'demo ', 'presentaci',
    ' día 1', ' día 2', ' día 3', ' dia 1', ' dia 2', ' dia 3',
    'bloque dia', 'bloque día',
})


CALENDAR_ID      = os.environ.get('CALENDAR_ID', '1c8aa3f1d9024d5eeead72447c0606f927cfd6ee6d1d2e5d28bf1b252959f396@group.calendar.google.com')
GOOGLE_API_KEY   = os.environ.get('GOOGLE_API_KEY', '')
GCAL_ICAL_URL    = os.environ.get('GCAL_ICAL_URL', '')   # iCal feed URL (no API key needed)
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

# ─── Auth helper ────────────────────────────────────────────────────────────

def _auth():
    return bool(session.get('compras_user') or session.get('cont_user'))

# ─── Shopify velocity ────────────────────────────────────────────────────────

# ─── Shopify velocity ────────────────────────────────────────────────────────

def _sync_shopify_orders(conn, days=60):
    """Wrapper · delega a shopify_client.sync_shopify_orders (helper unificado).

    Sebastián 23-may-2026 PM · consolidación 4→1 implementación · este
    archivo solo expone la firma legacy {ok, synced, days, error} que
    callers como `prog_sync_ventas` esperan.
    """
    from shopify_client import sync_shopify_orders as _sso
    return _sso(conn, days=days, incluir_movimientos=False)


def _shopify_velocity(conn, days=60):
    """
    Lee animus_shopify_orders de los últimos `days` días.
    Retorna dict {sku_full: vel_mes} y dict {producto_nombre: vel_mes}.

    Lookup de producto: exacto primero, luego parte antes del primer guión
    (pero NUNCA trunca a 6 chars — eso rompía SVITC33 y RECN-2).
    """
    since = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    # FIX 1-jun-2026 (audit): excluir canceladas/reembolsadas (inflan la velocidad →
    # sobreproducción). Mismo filtro que plan.py/auto_plan. Fallback si faltan columnas.
    _vel_base = ("SELECT sku_items, unidades_total, creado_en FROM animus_shopify_orders "
                 "WHERE creado_en >= ? AND sku_items IS NOT NULL")
    _vel_filtro = (" AND LOWER(COALESCE(estado,'')) NOT IN ('cancelled','cancelado','voided') "
                   "AND LOWER(COALESCE(estado_pago,'')) NOT IN "
                   "('refunded','voided','partially_refunded')")
    try:
        rows = conn.execute(_vel_base + _vel_filtro, (since,)).fetchall()
    except Exception:
        rows = conn.execute(_vel_base, (since,)).fetchall()

    sku_units = {}  # full_sku -> total units in period

    for row in rows:
        try:
            items = json.loads(row[0]) if row[0] else []
        except Exception:
            items = []
        for item in items:
            raw_sku = str(item.get('sku', '') or '').strip().upper()
            qty = int(item.get('qty', 0) or 0)
            if not raw_sku or qty <= 0:
                continue
            # Store full SKU — no truncation
            sku_units[raw_sku] = sku_units.get(raw_sku, 0) + qty

    # Denominador inteligente: prefiere ventana real solicitada (`days`) si
    # hay cobertura suficiente; sino cae a `actual_days` PERO marca data
    # quality como baja para que la UI advierta al usuario.
    #
    # Antes (CRITICAL fix #6 auditoría): siempre dividía por (d_max - d_min)
    # → si solo había 4 días de datos, la velocidad salía 10× más alta de
    # lo real, sobreestimando ventas y subestimando días de cobertura.
    actual_days = days  # default conservador
    data_quality = 'ok'
    coverage_pct = 100
    if rows:
        all_dates = [r[2] for r in rows if r[2]]
        if len(all_dates) >= 2:
            from datetime import date as _dt
            d_min = _dt.fromisoformat(min(all_dates)[:10])
            d_max = _dt.fromisoformat(max(all_dates)[:10])
            real_span = max((d_max - d_min).days, 1)
            coverage_pct = int(min(100, (real_span / max(days, 1)) * 100))
            # Si cobertura >= 80%: el período pedido es representativo,
            # dividir por `days` evita inflar por gaps al inicio/final.
            # Si < 80%: dividimos por real_span pero marcamos baja calidad.
            if coverage_pct >= 80:
                actual_days = days
                data_quality = 'ok'
            else:
                actual_days = real_span
                data_quality = 'low'  # poca data → no confiable para decisiones
        else:
            # 1 sola fecha: extrapolar es engañoso. Usamos days completo.
            actual_days = days
            data_quality = 'very_low'
            coverage_pct = 0
    else:
        actual_days = days
        data_quality = 'no_data'
        coverage_pct = 0
    months = max(actual_days / 30.0, 1/30.0)
    sku_vel = {sku: round(units / months, 1) for sku, units in sku_units.items()}

    # Build lookup map: SKU -> producto_nombre
    sku_map = {}
    for row in conn.execute("SELECT sku, producto_nombre FROM sku_producto_map WHERE activo=1").fetchall():
        sku_map[row[0].upper()] = row[1]

    prod_vel = {}  # producto_nombre -> vel_mes
    for sku, vel in sku_vel.items():
        # 1) Exact match (handles RECN-2, SVITC33, SVITC3315, NPHA10, etc.)
        prod = sku_map.get(sku)
        # 2) Fallback: try the part before the first hyphen (e.g. LBHA-30 -> LBHA)
        if not prod and '-' in sku:
            prod = sku_map.get(sku.split('-')[0])
        if prod:
            prod_vel[prod] = round(prod_vel.get(prod, 0) + vel, 1)

    return {
        'sku_velocity': sku_vel,
        'prod_velocity': prod_vel,
        'total_orders': len(rows),
        'days_requested': days,
        'actual_days_data': actual_days,
        'months_analyzed': round(months, 2),
        'data_quality': data_quality,    # ok|low|very_low|no_data
        'coverage_pct': coverage_pct,
    }

# ─── Google Calendar ─────────────────────────────────────────────────────────

def _parse_ical(text, days_ahead=90):
    """Parse iCal text and return list of {titulo, fecha, descripcion, id} dicts.

    Supports:
      - SUMMARY (titulo)
      - DESCRIPTION (descripcion, with line continuations)
      - DTSTART (start date)
      - UID (event id)
      - RRULE expansion: FREQ=DAILY/WEEKLY with INTERVAL and COUNT/UNTIL
        (covers our use case: recurrent batches every N days)
      - EXDATE (excluded dates)
      - Status: skip CANCELLED events
    """
    import re as _re
    today = date.today()
    cutoff = today + timedelta(days=days_ahead)
    # Window also includes some past so /auditoria-calendar can see history
    window_start = today - timedelta(days=180)
    events = []

    # Unfold lines: iCal continuation = next line starts with space/tab
    unfolded_lines = []
    for raw_line in text.splitlines():
        if raw_line.startswith((' ', '\t')) and unfolded_lines:
            unfolded_lines[-1] += raw_line[1:]
        else:
            unfolded_lines.append(raw_line)
    text_unfolded = '\n'.join(unfolded_lines)

    def _unescape_ical(s):
        return s.replace(r'\n', '\n').replace(r'\,', ',').replace(r'\;', ';').replace(r'\\', '\\')

    # Split into VEVENT blocks
    for block in _re.split(r'BEGIN:VEVENT', text_unfolded)[1:]:
        summary = ''
        description = ''
        dt_str = ''
        uid = ''
        status = ''
        rrule = ''
        exdates = set()
        for line in block.splitlines():
            line = line.rstrip()
            if line.startswith('SUMMARY:'):
                summary = _unescape_ical(line[8:]).strip()
            elif line.startswith('DESCRIPTION:'):
                description = _unescape_ical(line[12:]).strip()
            elif line.startswith('DTSTART'):
                val = line.split(':', 1)[-1].strip()
                dt_str = val[:8]
            elif line.startswith('UID:'):
                uid = line[4:].strip()
            elif line.startswith('STATUS:'):
                status = line[7:].strip().upper()
            elif line.startswith('RRULE:'):
                rrule = line[6:].strip()
            elif line.startswith('EXDATE'):
                # EXDATE;VALUE=DATE:20260518 or EXDATE:20260518T...
                val = line.split(':', 1)[-1].strip()
                for d_str in val.split(','):
                    d_str = d_str.strip()[:8]
                    if len(d_str) == 8:
                        try:
                            exdates.add(date(int(d_str[:4]), int(d_str[4:6]), int(d_str[6:8])))
                        except ValueError:
                            pass

        if not summary or len(dt_str) != 8 or status == 'CANCELLED':
            continue
        try:
            base_date = date(int(dt_str[:4]), int(dt_str[4:6]), int(dt_str[6:8]))
        except ValueError:
            continue

        # Generate occurrences
        occurrences = []
        if rrule:
            # Parse RRULE params
            rr = {}
            for part in rrule.split(';'):
                if '=' in part:
                    k, v = part.split('=', 1)
                    rr[k.strip().upper()] = v.strip()
            freq = rr.get('FREQ', '').upper()
            try:
                interval = int(rr.get('INTERVAL', '1'))
            except ValueError:
                interval = 1
            count = None
            if 'COUNT' in rr:
                try:
                    count = int(rr['COUNT'])
                except ValueError:
                    count = None
            until = None
            if 'UNTIL' in rr:
                u = rr['UNTIL'][:8]
                if len(u) == 8:
                    try:
                        until = date(int(u[:4]), int(u[4:6]), int(u[6:8]))
                    except ValueError:
                        until = None

            # Step in days
            if freq == 'DAILY':
                step_days = interval
            elif freq == 'WEEKLY':
                step_days = interval * 7
            elif freq == 'MONTHLY':
                step_days = interval * 30  # approx (good enough for our cadences)
            elif freq == 'YEARLY':
                step_days = interval * 365
            else:
                step_days = None

            if step_days:
                # Cap occurrences (safety: max 200 per series)
                max_occ = count if count is not None else 200
                cur = base_date
                for i in range(min(max_occ, 200)):
                    if until and cur > until:
                        break
                    if cur not in exdates:
                        occurrences.append(cur)
                    cur = cur + timedelta(days=step_days)
            else:
                occurrences = [base_date]
        else:
            occurrences = [base_date]

        # Filter by window and append
        for occ in occurrences:
            if window_start <= occ <= cutoff:
                events.append({
                    'titulo': summary,
                    'fecha': occ.isoformat(),
                    'descripcion': description,
                    'id': uid,
                })

    return sorted(events, key=lambda e: e['fecha'])


def _fetch_calendar_events(days_ahead=90):
    """Fetch production calendar events.
    Priority: 1) iCal feed (GCAL_ICAL_URL), 2) Google Calendar API (GOOGLE_API_KEY).
    """
    # ── Option 1: iCal feed (public or secret URL, no API key required) ──────
    if GCAL_ICAL_URL:
        try:
            req = urllib.request.Request(
                GCAL_ICAL_URL,
                headers={'User-Agent': 'EspagiRIA-Inventarios/1.0'}
            )
            with urllib.request.urlopen(req, timeout=10) as r:
                text = r.read().decode('utf-8', errors='replace')
            events = _parse_ical(text, days_ahead)
            return {'events': events, 'error': None, 'source': 'ical'}
        except Exception as e:
            # Sebastián 1-may-2026 audit: ANTES era silencioso · si Calendar
            # cae nadie sabía. Ahora log.warning + tipo de error explícito.
            log = logging.getLogger('inventario.programacion')
            log.warning('iCal fetch fallo (%s): %s', type(e).__name__, e)
            return {'events': [], 'error': f'iCal error: {e}', 'source': 'ical'}

    # ── Option 2: Google Calendar API (requires GOOGLE_API_KEY) ──────────────
    if not GOOGLE_API_KEY:
        return {'events': [], 'error': 'Configura GCAL_ICAL_URL en Render para leer el calendario', 'source': 'none'}

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    # Sebastián 1-may-2026: 'lunes vacío' · cambiar time_min al LUNES de esta
    # semana (no a now) para incluir eventos de lun-mar-mié-jue cuando hoy es vie.
    lunes_semana = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0)
    time_min = lunes_semana.strftime('%Y-%m-%dT%H:%M:%SZ')
    time_max = (now + timedelta(days=days_ahead)).strftime('%Y-%m-%dT%H:%M:%SZ')
    params = urllib.parse.urlencode({
        'key': GOOGLE_API_KEY,
        'timeMin': time_min,
        'timeMax': time_max,
        'singleEvents': 'true',
        'orderBy': 'startTime',
        'maxResults': 100,
    })
    cal_id_encoded = urllib.parse.quote(CALENDAR_ID, safe='')
    url = f"https://www.googleapis.com/calendar/v3/calendars/{cal_id_encoded}/events?{params}"
    try:
        req = urllib.request.Request(url, headers={'Accept': 'application/json'})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
        events = []
        for item in data.get('items', []):
            start = item.get('start', {})
            start_dt = start.get('dateTime', start.get('date', ''))[:10]
            events.append({
                'titulo': item.get('summary', 'Producción'),
                'fecha': start_dt,
                'descripcion': item.get('description', ''),
                'id': item.get('id', ''),
            })
        return {'events': events, 'error': None, 'source': 'gcal_api'}
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')[:300]
        return {'events': [], 'error': f'Calendar API HTTP {e.code}: {body}', 'source': 'gcal_api'}
    except Exception as e:
        return {'events': [], 'error': str(e), 'source': 'gcal_api'}

def _fetch_local_production_events(conn):
    """Return upcoming AND recent past (≤14d) production events from produccion_programada.

    Including recent past events lets the dashboard surface productions that were
    scheduled but not yet marked as 'completado', so they don't silently disappear
    from the Prox. Produccion column. Past-dated events carry past_date=True.
    """
    try:
        rows = conn.execute(
            """SELECT id, producto, fecha_programada, lotes, estado, observaciones
               FROM produccion_programada
               WHERE estado NOT IN ('completado','cancelado')
                 AND fecha_programada >= date('now', '-5 hours', '-14 days')
               ORDER BY fecha_programada"""
        ).fetchall()
        import datetime as _dt
        today_str = str(_dt.date.today())
        return [
            {'id': r[0], 'titulo': r[1], 'fecha': r[2],
             'lotes': r[3], 'estado': r[4], 'descripcion': r[5] or '',
             'past_date': r[2] < today_str}
            for r in rows
        ]
    except Exception as e:
        return []

# ─── MP stock ────────────────────────────────────────────────────────────────

def _norm_mp_name(name):
    """Normalise an MP name for fuzzy matching across two legacy ID systems.
    Removes: accents, parenthetical suffixes, hyphens->space, collapse spaces,
    and adds space between digits and letters (50KD -> 50 KD).
    """
    import unicodedata as _ud
    import re as _re
    s = str(name or '').strip()
    # Strip embedded control characters (Excel import artifacts)
    import re as _re2
    s = _re2.sub(r'[\x00-\x1f\x7f]', '', s)
    # Strip accents
    s = ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn')
    s = s.upper()
    # Remove parenthetical additions like (LYPHAR), (BASF), (1%)
    s = _re.sub(r'\([^)]*\)', '', s)
    # Hyphens and slashes to space
    s = _re.sub(r'[-/]', ' ', s)
    # Add space between digit and letter: 50KD -> 50 KD, 300KD -> 300 KD
    s = _re.sub(r'(\d)([A-Z])', r'\1 \2', s)
    s = _re.sub(r'([A-Z])(\d)', r'\1 \2', s)
    # Collapse multiple spaces
    s = _re.sub(r'\s+', ' ', s).strip()
    return s


def _lookup_stock_5tier(stock_mp, mid, nombre):
    """Lookup canónico de stock en el dict de _get_mp_stock con los 5 tiers
    (id → id.upper → nombre exacto → nombre normalizado → alias). Helper reusable ·
    Sebastián 1-jun-2026: varios lookups usaban menos tiers → 'Hay 0g' con stock /
    déficit falso (sobre-compra). Devuelve gramos (0 si no se encuentra)."""
    s = stock_mp.get(mid)
    if s is not None:
        return float(s)
    s = stock_mp.get((mid or '').upper())
    if s is not None:
        return float(s)
    _ne = (nombre or '').upper()
    s = stock_mp.get(_ne)
    if s is not None:
        return float(s)
    _nn = _norm_mp_name(nombre or '')
    s = stock_mp.get(_nn)
    if s is not None:
        return float(s)
    _al = _MP_NAME_ALIAS.get(_nn) or _MP_NAME_ALIAS.get(_ne)
    if _al and stock_mp.get(_al) is not None:
        return float(stock_mp.get(_al))
    return 0.0


# Static alias map: normalised formula name -> normalised movimientos name
# Handles brand-name differences, typos, Spanish/English variants, presentation suffixes
_MP_NAME_ALIAS = {
    # HIALURONICO missing ACIDO prefix
    'HIALURONICO 50 KD':   'ACIDO HIALURONICO 50 KD',
    'HIALURONICO 300 KD':  'ACIDO HIALURONICO 300 KD',
    'HIALURONICO 1500 KD': 'ACIDO HIALURONICO 1500 KD',
    # Typos in formula names
    'ACIDO HILAURONICO 50 KD':  'ACIDO HIALURONICO 50 KD',   # double A
    'ALCOHOL CETITLICO':        'ALCOHOL CETILICO',            # extra T
    'ACETYL TETRAPETIDE 5':     'ACETYL TETRAPEPTIDE 5',       # missing P
    'ACETYL TETRAPETIDE-5':     'ACETYL TETRAPEPTIDE 5',
    # Spanish -> English ingredient names
    'ACETIL TETRAPEPTIDO 40':  'ACETYL TETRAPEPTIDE 40',
    'ACETIL TETRAPEPTIDO-40':  'ACETYL TETRAPEPTIDE 40',
    'ACETIL TETRAPEPTIDO 3':   'ACETYL TETRAPEPTIDE 3',
    'ACETIL TETRAPEPTIDO-3':   'ACETYL TETRAPEPTIDE 3',
    # Brand/commercial name differences
    'ACEITE DE ARGAN':       'BEAUTY OIL ARGAN',
    'ACEITE ARGAN':          'BEAUTY OIL ARGAN',
    'ACEITE DE JOJOBA':      'BEAUTY OIL JOJOBA',
    'ACEITE JOJOBA':         'BEAUTY OIL JOJOBA',
    'ACEITE DE ROSA MOSQUETA': 'BEAUTY OIL ROSA MOSQUETA',
    # Presentation/form differences
    'ALOE VERA':             'ALOE VERA POLVO',
    'ACEITE ARBOL DE TE':    'ACEITE ESENCIAL ARBOL DE TE',
    'D PANTENOL':            'D PANTENOL LIQUIDO',
    'D-PANTENOL':            'D PANTENOL LIQUIDO',
    # Ascorbic acid derivatives
    '3 O ACIDO ETIL ASCORBICO': 'ETIL ASCORBICO ACID',
    'ASCORBIL GLUCOSIDE':        'ASCORBIL GLUCOSIDE',
    # Pantenol variants (without D- prefix)
    'PANTENOL LIQUIDO':   'D PANTENOL LIQUIDO',
    'PANTENOL SOLIDO':    'D PANTENOL SOLIDO',
    'PANTENOL':           'PANTENOL POLVO',
    # Centella variants
    'CENTELLA ASIATICA POLVO': 'CENTELLA ASIATICA',
    'CENTELLA ASIATICA':       'CENTELLA ASIATICA',
    'CENTELLA':                'CENTELLA ASIATICA',
    # Commercial name differences (formula uses generic, bodega uses brand)
    'EZ 4U':                   'PEMULEN EZ 4U',   # EZ-4U -> Pemulen EZ-4U
    'EZ-4U':                   'PEMULEN EZ 4U',
    'GRANSIL V 419':           'GRANSIL VX 419',   # Gransil variants
    'GRANSIL VX419':           'GRANSIL VX 419',
    'GRANSIL VX-419':          'GRANSIL VX 419',   # FIX 1-jun-2026 · era placeholder
    # 'PEMULEN EZ 4U' (falso match · descontaba MP equivocada vía tier-3 nombre)
    # Regaliz / licorice
    'REGALIZ':                 'EXTRACTO DE REGALIZ',
    'EXTRACTO REGALIZ':        'EXTRACTO DE REGALIZ',
    # Silicona (update once Bodega MP name confirmed)
    'SILICONA LIQUIDA':        'DIMETHICONE',
    'SILICONA':                'DIMETHICONE',
}


# MPs that are always available (own production equipment — effectively infinite stock)
_MP_UNLIMITED = {
    'AGUA DESIONIZADA', 'AGUA PURIFICADA', 'AGUA DESTILADA',
    'AGUA PURIFICADA TOTAL', 'AGUA', 'AQUA',
}
_MP_UNLIMITED_NORM = set()  # populated lazily in _get_mp_stock


def _detect_alias_collisions(conn):
    """Detecta nombres de MP distintos que colapsan al mismo nombre normalizado.

    Si dos MPs reales con codigos distintos normalizan al mismo nombre, el
    alias map puede hacer que el stock de uno se atribuya al otro silenciosamente.
    Esta función lo detecta y devuelve la lista para alertar al usuario.

    Returns: lista de dicts con {'norm', 'variantes': [{codigo, nombre}, ...]}
    """
    by_norm = {}
    try:
        rows = conn.execute(
            "SELECT codigo_mp, nombre_comercial FROM maestro_mps WHERE activo=1"
        ).fetchall()
        for codigo, nombre in rows:
            n = _norm_mp_name(nombre or '')
            if not n:
                continue
            by_norm.setdefault(n, []).append({
                'codigo': str(codigo or ''),
                'nombre': str(nombre or ''),
            })
    except Exception:
        return []
    collisions = []
    for n, variantes in by_norm.items():
        if len(variantes) >= 2:
            # Solo es colisión real si los nombres originales son distintos
            distintos = {v['nombre'].strip().upper() for v in variantes}
            if len(distintos) >= 2:
                collisions.append({'normalizado': n, 'variantes': variantes})
    return collisions


def _is_unlimited_mp(nombre):
    """Return True if the MP is produced on-site (water equipment, etc.)."""
    n = str(nombre or '').strip().upper()
    return n in _MP_UNLIMITED or any(u in n for u in ('AGUA DESIONIZADA', 'AGUA PURIFICADA'))


def _get_mp_stock(conn):
    """Returns dict {key: stock_actual_g} keyed by material_id AND all known nombres.

    Two-pass strategy to avoid split-stock bugs when the same material_id appears
    in movimientos under different material_nombre values (e.g. 'PROPYLENE GLYCOL'
    and 'PROPILENGLICOL' are both MP00121):

    Pass 1 – aggregate stock correctly by material_id (canonical total).
    Pass 2 – collect all (material_id, material_nombre) pairs; map every name
             to the canonical stock for that material_id.
    """
    # PERF-FIX 23-may-2026 · auditoría · función pesada (3 full-scans sobre
    # movimientos) llamada 11+ veces por request en dashboards de planta ·
    # memoizar en flask.g por request (válido solo durante el request actual)
    try:
        cached = getattr(g, '_mp_stock_cache', None)
        if cached is not None:
            return cached
    except RuntimeError:
        # Sin contexto de request (cron job) · no memoizar
        pass
    # INVIMA-FIX · 22-may-2026 · ABASTECIMIENTO ZERO-ERROR (#2 + #3 audit 22-may)
    # · ANTES: SUM no excluía CUARENTENA/RECHAZADO/VENCIDO → planificación
    #   creía stock disponible y NO recomendaba compra → paro producción.
    # · ANTES: Ajuste/Ajuste+ contaba como Salida (bug suma) → over-ordering.
    # · AHORA: Entrada+Ajuste+Ajuste+ suman · Salida+Ajuste- restan · excluye
    #   estados no-disponibles · consistente con inventario_helpers canonical.
    # Pass 1: canonical stock per material_id
    id_stock = {}
    for mid, sg in conn.execute("""
        SELECT material_id,
               COALESCE(SUM(
                 CASE
                   WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                   WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                   ELSE 0
                 END), 0)
        FROM movimientos
        WHERE material_id IS NOT NULL AND material_id != ''
          AND (estado_lote IS NULL
               OR UPPER(COALESCE(estado_lote,'')) NOT IN
                  ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))
        GROUP BY material_id
    """).fetchall():
        id_stock[str(mid).strip()] = max(float(sg or 0), 0)

    # Pass 2: collect all name variants per material_id
    name_rows = conn.execute(
        "SELECT DISTINCT material_id, material_nombre FROM movimientos "
        "WHERE material_nombre IS NOT NULL AND material_nombre != ''"
    ).fetchall()

    stock = {}
    # Index by canonical material_id first
    for mid, val in id_stock.items():
        stock[mid] = val

    # Index by every name variant, resolved to the canonical material_id stock
    for mid, nombre in name_rows:
        mid_key = str(mid or '').strip()
        val = id_stock.get(mid_key, 0)  # canonical stock for this material_id
        nombre_s = str(nombre).strip()
        # exact uppercase
        key_exact = nombre_s.upper()
        if key_exact not in stock:
            stock[key_exact] = val
        # normalised (strips accents, hyphens, parentheses, control chars)
        key_norm = _norm_mp_name(nombre_s)
        if key_norm and key_norm not in stock:
            stock[key_norm] = val

    # Also index materials that only appear without a material_id (legacy rows)
    # ABASTECIMIENTO-FIX · 22-may-2026 · misma corrección Ajuste/CUARENTENA
    for nombre, sg in conn.execute("""
        SELECT material_nombre,
               COALESCE(SUM(
                 CASE
                   WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                   WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                   ELSE 0
                 END), 0)
        FROM movimientos
        WHERE (material_id IS NULL OR material_id = '')
          AND material_nombre IS NOT NULL AND material_nombre != ''
          AND (estado_lote IS NULL
               OR UPPER(COALESCE(estado_lote,'')) NOT IN
                  ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))
        GROUP BY material_nombre
    """).fetchall():
        val = max(float(sg or 0), 0)
        key_exact = str(nombre).strip().upper()
        if key_exact not in stock:
            stock[key_exact] = val
        key_norm = _norm_mp_name(str(nombre))
        if key_norm and key_norm not in stock:
            stock[key_norm] = val

    # Bridge tier (tier 5): formula_material_id → bodega_material_id → canonical stock.
    # This resolves cases where formula_items uses one ID system and movimientos uses another.
    # The mp_formula_bridge table maps them explicitly (populated via admin UI).
    try:
        bridge_rows = conn.execute(
            "SELECT formula_material_id, bodega_material_id FROM mp_formula_bridge WHERE activo=1"
        ).fetchall()
        for fid, bid in bridge_rows:
            fid_key = str(fid or '').strip()
            bid_key = str(bid or '').strip()
            if not fid_key or not bid_key:
                continue
            bodega_stock = id_stock.get(bid_key, 0)
            # Index by formula_material_id (primary bridge key)
            if fid_key not in stock:
                stock[fid_key] = bodega_stock
            # Also index by normalised formula name if available (from bridge table)
    except Exception:
        pass  # bridge table may not exist in older DB snapshots

    # PERF-FIX 23-may-2026 · memoizar para el resto del request
    try:
        g._mp_stock_cache = stock
    except RuntimeError:
        pass  # sin contexto de request
    return stock

# ─── Formula lookup ──────────────────────────────────────────────────────────

def _get_formulas(conn):
    """Returns dict {producto_nombre: {'lote_size_kg': X, 'items': [...{material_id,pct,g}...]}}"""
    headers = conn.execute(
        "SELECT producto_nombre, unidad_base_g, lote_size_kg FROM formula_headers"
    ).fetchall()
    items_all = conn.execute(
        "SELECT producto_nombre, material_id, material_nombre, porcentaje, cantidad_g_por_lote FROM formula_items"
    ).fetchall()

    items_map = {}
    for row in items_all:
        p = row[0]
        items_map.setdefault(p, []).append({
            'material_id': row[1],
            'material_nombre': row[2],
            'porcentaje': float(row[3] or 0),
            'cantidad_g_por_lote': float(row[4] or 0),
        })

    formulas = {}
    for h in headers:
        pname = h[0]
        ub_g = float(h[1] or 0)
        lote_kg = float(h[2] or (ub_g / 1000 if ub_g else 0))
        formulas[pname] = {
            'lote_size_kg': lote_kg,
            'unidad_base_g': ub_g,
            'items': items_map.get(pname, []),
        }
    return formulas

# ─── Helpers: PT stock y MEE stock ──────────────────────────────────────────

def _resolved_stock_por_sku(conn, empresa=None):
    """Stock disponible por SKU aplicando regla 'CC manda sobre SHOPIFY'.

    Sebastián 12-may-2026: extraído de _get_stock_pt para reuso desde el
    panel de prioridad-agotamiento (admin.py). Antes ese panel hacía un
    SUM directo que doble-contaba CC + SHOPIFY · ahora usa esta función.

    REGLA DE AUTORIDAD: si para un SKU hay rows liberados por CC (lote
    NO empieza con 'SHOPIFY-'), ESOS son la fuente de verdad y se
    IGNORAN los snapshots Shopify para ese SKU (la app local ya descuenta
    al vender; sumar Shopify duplicaba). Si SOLO hay rows SHOPIFY (SKU
    sin pasar por maquila local), se usa Shopify como fallback.

    Args:
        conn: conexión sqlite
        empresa: si se pasa (ej. 'ANIMUS'), filtra solo stock de esa empresa.

    Returns:
        dict {sku_upper: {'uds': int, 'descripcion': str, 'fuente': 'CC'|'SHOPIFY'}}
    """
    cc_stock = {}      # SKU → uds (autoridad)
    shop_stock = {}    # SKU → uds (fallback)
    descripcion = {}   # SKU → descripcion (la más reciente)
    shop_max_age_hours = 0.0

    where_empresa = " AND UPPER(TRIM(COALESCE(empresa,''))) = UPPER(?)" if empresa else ""
    params = (empresa,) if empresa else ()

    try:
        rows = conn.execute(f"""
            SELECT UPPER(TRIM(sku)) AS sku,
                   MAX(COALESCE(lote_produccion, '')) AS lote,
                   COALESCE(SUM(unidades_disponible), 0) AS uds,
                   MAX(COALESCE(descripcion, '')) AS descripcion,
                   MAX(COALESCE(fecha_liberacion, fecha_creacion, '')) AS fmax
            FROM stock_pt
            WHERE estado = 'Disponible' {where_empresa}
            GROUP BY UPPER(TRIM(sku)),
                     CASE WHEN COALESCE(lote_produccion,'') LIKE 'SHOPIFY-%'
                          THEN 'SHOPIFY' ELSE 'CC' END
        """, params).fetchall()
    except sqlite3.OperationalError:
        rows = conn.execute(f"""
            SELECT UPPER(TRIM(sku)) AS sku,
                   MAX(COALESCE(lote_produccion, '')) AS lote,
                   COALESCE(SUM(unidades_disponible), 0) AS uds,
                   MAX(COALESCE(descripcion, '')) AS descripcion,
                   '' AS fmax
            FROM stock_pt
            WHERE estado = 'Disponible' {where_empresa}
            GROUP BY UPPER(TRIM(sku)),
                     CASE WHEN COALESCE(lote_produccion,'') LIKE 'SHOPIFY-%'
                          THEN 'SHOPIFY' ELSE 'CC' END
        """, params).fetchall()

    from datetime import datetime as _dt
    for row in rows:
        raw_sku = str(row[0] or '').strip().upper()
        lote = str(row[1] or '')
        uds = max(int(row[2] or 0), 0)
        desc = str(row[3] or '')
        fmax = str(row[4] or '')
        if not raw_sku or uds <= 0:
            continue
        if desc and not descripcion.get(raw_sku):
            descripcion[raw_sku] = desc
        if lote.startswith('SHOPIFY-'):
            shop_stock[raw_sku] = shop_stock.get(raw_sku, 0) + uds
            if fmax:
                try:
                    age_h = (_dt.now(timezone.utc).replace(tzinfo=None) - _dt.fromisoformat(fmax.replace('Z', ''))).total_seconds() / 3600.0
                    if age_h > shop_max_age_hours:
                        shop_max_age_hours = age_h
                except Exception:
                    pass
        else:
            cc_stock[raw_sku] = cc_stock.get(raw_sku, 0) + uds

    if shop_max_age_hours > 24:
        logging.getLogger('programacion').warning(
            "Stock PT desde SHOPIFY tiene %.1fh de antigüedad — revisa el cron de sync",
            shop_max_age_hours
        )

    resolved = {}
    for sku, uds in cc_stock.items():
        resolved[sku] = {'uds': uds, 'descripcion': descripcion.get(sku, ''), 'fuente': 'CC'}
    for sku, uds in shop_stock.items():
        if sku not in resolved:
            resolved[sku] = {'uds': uds, 'descripcion': descripcion.get(sku, ''), 'fuente': 'SHOPIFY'}
    return resolved


def _get_stock_pt(conn):
    """Stock real de producto terminado desde stock_pt (por producto_nombre).

    REGLA DE AUTORIDAD (fix CRITICAL #1 auditoría): si para un SKU hay rows
    liberados por CC (lote_produccion NO empieza con 'SHOPIFY-'), ESOS son
    la fuente de verdad. Los snapshots de Shopify se IGNORAN para ese SKU
    (porque la app ya descuenta y libera; los rows SHOPIFY son redundantes
    y al sumarlos se doble contaba el stock).

    Si para un SKU SOLO hay rows de SHOPIFY (porque nunca se ha liberado
    desde CC), se usan como fallback. Eso cubre productos que aún no han
    pasado por el flujo de maquila local.

    Loguea warning si los rows SHOPIFY tienen >24h (staleness) — señal de
    que el sync se está retrasando y conviene revisar el cron.

    Sebastián 12-may-2026: refactor · ahora delega a _resolved_stock_por_sku
    para que panel admin y este coincidan en la regla.
    """
    sku_map = {}
    try:
        for row in conn.execute(
            "SELECT sku, producto_nombre FROM sku_producto_map WHERE activo=1"
        ).fetchall():
            k = str(row[0] or '').strip().upper()
            if k:
                sku_map[k] = str(row[1] or '').strip().upper()
    except Exception:
        pass

    resolved = _resolved_stock_por_sku(conn)

    # Mapear SKU → producto_nombre (agregado para back-compat)
    stock = {}
    for raw_sku, info in resolved.items():
        uds = info['uds']
        prod = sku_map.get(raw_sku)
        if not prod:
            prefix = raw_sku.split('-')[0]
            prod = sku_map.get(prefix)
        if prod:
            stock[prod] = stock.get(prod, 0) + uds
        else:
            stock[raw_sku] = stock.get(raw_sku, 0) + uds

    return stock


def _get_mee_stock(conn):
    """
    MEE stock CANONICAL from movimientos_mee (Entrada-Salida).
    Falls back to maestro_mee.stock_actual where no movements exist
    (defensa para MEEs viejos sin kardex completo · drift mitigado por
    cron mee_drift_sync diario 3 AM).
    Returns dict {codigo_upper: stock_float}

    AUDITORÍA-FIX 23-may-2026 · C18 · memoizado en flask.g por request
    (mismo patrón que _get_mp_stock) · evita re-scan de movimientos_mee
    en hot paths con múltiples lookups
    """
    try:
        cached = getattr(g, '_mee_stock_cache', None)
        if cached is not None:
            return cached
    except RuntimeError:
        pass
    # From movements (accurate)
    # P0-3 23-may-PM · auditoría agente Stock · antes solo Entrada/Salida ·
    # ignoraba 'Ajuste' (signed positive) + variantes case mixto que SÍ
    # existen en BD · y job_mee_drift_sync los cuenta, causando drift
    # permanente entre cron y canónico. Ahora unifica con _get_mp_stock.
    stock = {}
    try:
        for row in conn.execute("""
            SELECT mee_codigo,
                   COALESCE(SUM(CASE
                       WHEN LOWER(tipo) IN ('entrada','ingreso','devolucion','devolución','ajuste')
                           THEN cantidad
                       WHEN LOWER(tipo) IN ('salida','consumo','rechazo')
                           THEN -cantidad
                       ELSE 0 END), 0)
            FROM movimientos_mee WHERE COALESCE(anulado,0)=0 GROUP BY mee_codigo
        """).fetchall():
            k = str(row[0] or '').strip().upper()
            if k:
                stock[k] = max(float(row[1] or 0), 0)
    except Exception:
        pass
    # Fill in maestro_mee.stock_actual for codes with no movements
    try:
        for row in conn.execute("SELECT codigo, stock_actual FROM maestro_mee").fetchall():
            k = str(row[0] or '').strip().upper()
            if k and k not in stock:
                stock[k] = max(float(row[1] or 0), 0)
    except Exception:
        pass
    try:
        g._mee_stock_cache = stock
    except RuntimeError:
        pass
    return stock


# ─── Stock projection ────────────────────────────────────────────────────────

# Umbrales de cobertura (días). Se leen de env si están definidos para
# permitir ajuste sin redeploy. Default 20/40 = comportamiento histórico.
def _dias_thresholds():
    try:
        c = max(1, int(os.environ.get('DIAS_CRITICOS') or 20))
        a = max(c + 1, int(os.environ.get('DIAS_ALERTA') or 40))
        return c, a
    except (ValueError, TypeError):
        return 20, 40

DIAS_CRITICOS, DIAS_ALERTA = _dias_thresholds()

# Lead time por MP (días). Si una MP viene de China, el sistema necesita
# anticipar la compra ~60 días antes del agotamiento. Para MPs locales,
# 21 días es razonable (compra → entrega).
LEAD_TIME_CHINA = int(os.environ.get('LEAD_TIME_CHINA') or 60)
LEAD_TIME_LOCAL = int(os.environ.get('LEAD_TIME_LOCAL') or 21)


def _compute_mp_deficit_aggregated(conn, days_ahead=90):
    """Calcula déficit REAL de MPs agregando primero, restando stock una sola vez.

    Esta es la lógica correcta — agrupa todas las producciones planificadas
    (calendario + DB local), suma cuánto se necesita por MP, y resta el stock
    actual UNA vez para obtener el déficit real.

    El bug que esto fixea: si calculas deficit per-product y sumas, cuando hay
    stock parcial cada producto "ve" el mismo stock disponible y under-cuentas
    el déficit total. Ej: A pide 3g, B pide 5g, stock=4g → A: déf=0, B: déf=1,
    suma=1g. Real: 3+5-4 = 4g. La suma per-product subestima en 3g.

    Returns dict {material_id: {nombre, total_g, stock_g, deficit_g, productos,
                                proveedor, por_mes, n_meses}}.
    Solo incluye MPs con deficit_g > 0. MPs ilimitados (agua, etc.) excluidos.
    """
    import datetime as _dt
    import re as _re

    cal = _fetch_calendar_events(days_ahead=days_ahead)
    events = cal.get('events', [])
    formulas = _get_formulas(conn)
    mp_stock = _get_mp_stock(conn)
    if not formulas:
        return {}

    # SKU → producto map
    _sku_to_prod = {}
    try:
        for row in conn.execute(
            "SELECT UPPER(sku), producto_nombre FROM sku_producto_map WHERE activo=1"
        ).fetchall():
            _sku_to_prod[row[0]] = row[1]
    except Exception:
        pass

    # Proveedor por MP · Sebastian 4-may-2026 (Catalina): normalizar al
    # primer "case canónico" para que "Agenquimicos" y "AGENQUIMICOS" no
    # caigan en grupos distintos al agrupar por proveedor.
    _prov_map = {}
    _prov_canonical = {}  # lowercase trimmed → primera variante observada
    try:
        for row in conn.execute(
            "SELECT codigo_mp, COALESCE(proveedor,'') FROM maestro_mps"
        ).fetchall():
            prov_raw = (row[1] or '').strip()
            if not prov_raw:
                _prov_map[row[0]] = ''
                continue
            key = prov_raw.lower()
            canonical = _prov_canonical.setdefault(key, prov_raw)
            _prov_map[row[0]] = canonical
    except Exception:
        pass

    today = _dt.date.today()
    cutoff = today + _dt.timedelta(days=days_ahead)

    # Tokens que no son SKUs en títulos del calendario
    _NOT_SKU = {
        'FAB','QC','CON','SIN','MICRO','ENVASADO','ACONDICIONAMIENTO',
        'FABRICACION','FABRICACIÓN','LANZAMIENTO','PRODUCCION','PRODUCCIÓN',
        'KG','MES','DIAS','DÍAS','ML','UDS','BATCH','FERNANDO',
        'MESA','LOTES','LOTE','MINI','MACRO','AND','THE','FOR'
    }

    def _skus(titulo):
        tokens = _re.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', titulo.upper())
        return [t for t in tokens if t not in _NOT_SKU]

    def _kg_ev(titulo):
        m = _re.findall(r'~?(\d+(?:[,.]\d+)*)\s*kg', titulo, _re.IGNORECASE)
        if not m:
            return None
        try:
            return float(m[-1].replace(',', '.'))
        except Exception:
            return None

    # Filtro: ignorar fases que NO consumen MPs (envasado/acondicionamiento/QC)
    # Sebastián 12-may-2026: usar constante global unificada.
    _NON_FAB_KW = NON_FAB_KW_GLOBAL

    producciones = []
    seen_events = set()

    # Calendario
    for ev in events:
        titulo = ev.get('titulo', '')
        fecha_s = ev.get('fecha', '')
        if not fecha_s:
            continue
        try:
            fecha = _dt.date.fromisoformat(fecha_s)
        except ValueError:
            continue
        if fecha < today or fecha > cutoff:
            continue
        if any(kw in titulo.lower() for kw in _NON_FAB_KW):
            continue
        key = (titulo.strip(), fecha_s)
        if key in seen_events:
            continue
        seen_events.add(key)
        kg = _kg_ev(titulo)
        for sku in _skus(titulo):
            prod = _sku_to_prod.get(sku)
            if prod and prod in formulas:
                producciones.append({
                    'fecha': fecha_s, 'mes': fecha.strftime('%Y-%m'),
                    'producto': prod, 'kg': kg or formulas[prod]['lote_size_kg'],
                })
                break

    # produccion_programada (DB local) — complementa el calendario
    _upper_to_prod = {p.upper(): p for p in formulas.keys()}
    try:
        local_rows = conn.execute(
            """SELECT producto, fecha_programada, lotes FROM produccion_programada
               WHERE estado NOT IN ('completado','cancelado')
                 AND fecha_programada >= date('now', '-5 hours', '-7 days')
                 AND fecha_programada <= ?
               ORDER BY fecha_programada""",
            (cutoff.isoformat(),)
        ).fetchall()
        for row in local_rows:
            prod_raw = (row[0] or '').strip()
            fecha_s = (row[1] or '').strip()
            lotes = int(row[2] or 1)
            prod = prod_raw if prod_raw in formulas else _upper_to_prod.get(prod_raw.upper())
            if not prod or not fecha_s:
                continue
            key = (prod, fecha_s)
            if key in seen_events:
                continue
            seen_events.add(key)
            try:
                fecha = _dt.date.fromisoformat(fecha_s)
            except ValueError:
                continue
            lote_kg = formulas[prod]['lote_size_kg']
            producciones.append({
                'fecha': fecha_s, 'mes': fecha.strftime('%Y-%m'),
                'producto': prod, 'kg': lote_kg * lotes,
            })
    except Exception:
        pass

    # Agregar necesidad por MP a través de TODAS las producciones
    mp_needed = {}  # mid → {nombre, total_g, productos, por_mes}
    for pr in producciones:
        prod = pr['producto']
        kg_ev = pr['kg']
        formula = formulas.get(prod, {})
        lote_kg = formula.get('lote_size_kg', 1)
        factor = kg_ev / lote_kg if lote_kg > 0 else 1.0
        for item in formula.get('items', []):
            nombre = item.get('material_nombre', '')
            # FIX 2-jun-2026 audit abastecimiento (M1) · resolver a código de BODEGA
            # antes de acumular demanda · igual que consumo_horizontes · evita demanda
            # partida entre códigos de fórmula del mismo material (este motor alimenta
            # /generar-oc, /regenerar-oc, /mps-deficit).
            _mid_raw = str(item['material_id']).strip()
            mid = _resolver_material_bodega(conn, _mid_raw, nombre) or _mid_raw
            g_lote = item.get('cantidad_g_por_lote', 0)
            g_need = float(g_lote) * factor
            if g_need <= 0 or not mid:
                continue
            if _is_unlimited_mp(nombre):
                continue
            if mid not in mp_needed:
                mp_needed[mid] = {
                    'nombre': nombre, 'total_g': 0.0,
                    'productos': [], 'por_mes': {},
                }
            mp_needed[mid]['total_g'] += g_need
            mp_needed[mid]['por_mes'][pr['mes']] = (
                mp_needed[mid]['por_mes'].get(pr['mes'], 0) + g_need
            )
            if prod not in mp_needed[mid]['productos']:
                mp_needed[mid]['productos'].append(prod)

    # Lookup stock por mid o por nombre normalizado
    # FIX 1-jun-2026 audit Abastecimiento (P0) · faltaba el tier de ALIAS que sí
    # tiene el lookup canónico (~1404) → MPs con nombre-variante no bridgeado
    # resolvían stock=0 → déficit FALSO → sobre-compra. Alineado al canónico.
    def _lookup_stock(mid, nombre):
        s = mp_stock.get(mid)
        if s is not None:
            return s
        s = mp_stock.get((mid or '').upper())
        if s is not None:
            return s
        _nom_exact = (nombre or '').upper()
        s = mp_stock.get(_nom_exact)
        if s is not None:
            return s
        _nom_norm = _norm_mp_name(nombre or '')
        s = mp_stock.get(_nom_norm)
        if s is not None:
            return s
        _alias = _MP_NAME_ALIAS.get(_nom_norm) or _MP_NAME_ALIAS.get(_nom_exact)
        if _alias and mp_stock.get(_alias) is not None:
            return mp_stock.get(_alias)
        return 0

    # Calcular déficit (UNA sola resta — total_g - stock)
    out = {}
    for mid, data in mp_needed.items():
        stock_g = _lookup_stock(mid, data['nombre'])
        deficit = max(0.0, data['total_g'] - stock_g)
        if deficit <= 0:
            continue
        out[mid] = {
            'material_id': mid,
            'nombre': data['nombre'],
            'total_g': round(data['total_g'], 1),
            'stock_g': round(stock_g, 1),
            'deficit_g': round(deficit, 1),
            'productos': data['productos'],
            'por_mes': data['por_mes'],
            'n_meses': len(data['por_mes']),
            'proveedor': _prov_map.get(mid, ''),
        }
    return out


def _project_stock(conn, prod_vel, formulas, mp_stock, calendar_events, china_mps=None):
    """
    Logica correcta de programacion:
    1. Stock PT real = producido (acondicionamiento) - vendido (Shopify)
    2. Vel diaria = vel_mes / 30
    3. Dias de cobertura = stock / vel_diaria
    4. Umbral ROJO < 20 dias, AMARILLO < 40 dias, VERDE >= 40 dias
    5. Validar calendario: hay produccion antes del dia critico?
    6. Validar MPs: alcanzan para un lote?
    7. Alertas de compra si faltan MPs o MEE
    """
    # Stock real de PT
    pt_stock = _get_stock_pt(conn)

    # Calendar: next production date per product
    # Priority: 1) local DB, 2) Google Calendar / iCal via SKU lookup
    products_with_formula = set(formulas.keys())
    next_prod_by_product = {}

    # 1. Local DB events — case-insensitive product name matching
    local_events = _fetch_local_production_events(conn)
    # Build uppercase lookup: UPPER(formula_name) -> original formula_name
    _upper_to_prod = {p.upper(): p for p in products_with_formula}
    for ev in local_events:
        prod_ev_raw = ev.get('titulo', '')
        # Try exact match first, then case-insensitive
        prod_key = prod_ev_raw if prod_ev_raw in products_with_formula             else _upper_to_prod.get(prod_ev_raw.upper())
        if prod_key and prod_key not in next_prod_by_product:
            fecha_ev = ev.get('fecha', '')
            next_prod_by_product[prod_key] = fecha_ev
            # Flag if the scheduled date is already past
            if ev.get('past_date'):
                next_prod_by_product[prod_key + '__past'] = True

    # 2. Google Calendar / iCal events — SKU-first matching
    # Event titles use SKU codes: 'GELH – Fabricacion' -> SKU=GELH -> GEL HIDRATANTE
    import re as _re_cal

    # Load sku -> producto_nombre map from DB
    _sku_to_prod = {}
    try:
        for row in conn.execute(
            "SELECT UPPER(sku), producto_nombre FROM sku_producto_map WHERE activo=1"
        ).fetchall():
            _sku_to_prod[row[0]] = row[1]
    except Exception:
        pass

    # Non-SKU words that appear in event titles — never treat as product codes
    _NOT_SKU = {
        'FAB','QC','CON','SIN','MICRO','ENVASADO','ACONDICIONAMIENTO',
        'FABRICACION','FABRICACIÓN','LANZAMIENTO','PRODUCCION','PRODUCCIÓN',
        'KG','MES','DIAS','DÍAS','ML','UDS','SIN','CON','BATCH','Fernando',
        'MESA','LOTES','LOTE','MINI','MACRO','AND','THE','FOR'
    }

    def _skus_from_title(titulo):
        """Extract SKU candidates from a calendar event title."""
        # All uppercase tokens 2+ chars that are not generic words
        tokens = _re_cal.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', titulo.upper())
        return [t for t in tokens if t not in _NOT_SKU]

    next_prod_kg_by_product = {}  # product -> planned kg from calendar title

    def _kg_from_title(titulo):
        """Extract planned kg from calendar event title. Returns float or None."""
        # Match patterns like '50 kg', '~29 kg', '196 kg', '1,250 u / 50 kg'
        matches = _re_cal.findall(r'~?(\d+(?:[,.]\d+)*)\s*kg', titulo, _re_cal.IGNORECASE)
        if not matches:
            return None
        # Take the last match (avoids 'u' counts like '1,250 u / 50 kg')
        val = matches[-1].replace(',', '.')
        try:
            return float(val)
        except ValueError:
            return None

    _today_str = str(date.today())

    for ev in calendar_events:
        titulo  = ev.get('titulo', '')
        fecha_ev = ev.get('fecha', '')
        if not fecha_ev:
            continue
        # Skip past calendar events — always show next FUTURE production date
        if fecha_ev < _today_str:
            continue
        # Extract planned kg from title
        kg_ev = _kg_from_title(titulo)
        # Try each SKU extracted from title
        for sku in _skus_from_title(titulo):
            prod_name = _sku_to_prod.get(sku)
            if prod_name and prod_name in products_with_formula:
                # Assign earliest FUTURE date per product
                existing = next_prod_by_product.get(prod_name)
                if not existing or fecha_ev < existing:
                    next_prod_by_product[prod_name] = fecha_ev
                    if kg_ev:
                        next_prod_kg_by_product[prod_name] = kg_ev
                break

    today = date.today()

    projection = []
    all_alerts = []

    for prod, formula in sorted(formulas.items()):
        vel_mes = prod_vel.get(prod, 0)
        vel_dia = vel_mes / 30.0 if vel_mes > 0 else 0.0
        lote_kg = formula['lote_size_kg']
        items   = formula['items']

        # Stock actual PT en unidades
        stock_uds = pt_stock.get(prod.upper(), 0)

        # Dias de cobertura
        # CRITICAL fix: si vel_dia=0 Y stock_uds=0, NO marcar 999 (eso hace
        # cal_ok siempre true y oculta faltantes reales). Mejor None y tratar
        # explícitamente abajo.
        if vel_dia > 0:
            dias_cob = stock_uds / vel_dia
            tiene_datos = True
        elif stock_uds > 0:
            dias_cob = 999.0   # hay stock pero sin ventas = cobertura indefinida real
            tiene_datos = True
        else:
            dias_cob = 0.0     # sin stock y sin ventas → cobertura cero, dato dudoso
            tiene_datos = False

        # Produccion en calendario
        prox_prod = next_prod_by_product.get(prod, 'No programado')

        # Validar si la produccion calendario llega antes del dia critico.
        # Si no hay datos (sin stock ni ventas), no asumir que llega "a tiempo".
        cal_ok = False
        if prox_prod != 'No programado' and tiene_datos:
            try:
                prod_date = date.fromisoformat(prox_prod)
                dias_hasta_prod = (prod_date - today).days
                # cal_ok solo si hay velocidad real Y la prod llega antes del agotamiento
                cal_ok = vel_dia > 0 and dias_hasta_prod <= dias_cob
            except Exception:
                cal_ok = False

        # MP check: alcanza la MP para el lote del calendario?
        # Scale formula quantities by planned kg (from calendar) vs reference lot
        planned_kg = next_prod_kg_by_product.get(prod, lote_kg or 1.0)
        ref_kg     = lote_kg if lote_kg and lote_kg > 0 else planned_kg
        kg_scale   = planned_kg / ref_kg  # e.g. 50kg planned / 35kg ref = 1.43x

        mp_check = []
        can_produce = True        # False only if MP found-in-system but insufficient
        has_data_gap = False      # True if any MP not found in movimientos at all
        items_with_qty = [i for i in items if i.get('cantidad_g_por_lote', 0) > 0]
        if not items_with_qty:
            can_produce = None  # formula sin cantidades definidas
        else:
            for item in items_with_qty:
                mid = str(item['material_id']).strip()
                needed_g = float(item['cantidad_g_por_lote']) * kg_scale
                nombre_raw = str(item.get('material_nombre', '')).strip()
                # Lookup order: unlimited → id → exact name → norm name → alias → NOT_FOUND
                mp_found = True
                if _is_unlimited_mp(nombre_raw):
                    available_g = float('inf')
                elif mid in mp_stock:
                    available_g = mp_stock[mid]
                else:
                    nombre_exact = nombre_raw.upper()
                    if nombre_exact in mp_stock:
                        available_g = mp_stock[nombre_exact]
                    else:
                        nombre_norm = _norm_mp_name(nombre_raw)
                        if nombre_norm in mp_stock:
                            available_g = mp_stock[nombre_norm]
                        else:
                            alias_key = _MP_NAME_ALIAS.get(nombre_norm) or _MP_NAME_ALIAS.get(nombre_exact)
                            if alias_key and alias_key in mp_stock:
                                available_g = mp_stock[alias_key]
                            else:
                                available_g = 0
                                mp_found = False  # genuinely not in movimientos
                deficit_g = 0 if available_g == float('inf') else max(0, needed_g - available_g)
                ok = deficit_g < 1
                if not ok:
                    if mp_found:
                        can_produce = False  # confirmed deficit: found but insufficient
                    else:
                        has_data_gap = True  # data gap: not in movimientos, unknown stock
                mp_check.append({
                    'material_id': mid,
                    'nombre': item['material_nombre'],
                    'needed_g': round(needed_g, 1),
                    'available_g': '∞' if available_g == float('inf') else available_g,
                    'deficit_g': round(deficit_g, 1),
                    'ok': ok,
                    'mp_found': mp_found,
                })

        missing_mp = [m for m in mp_check if not m['ok']]

        # Semaforo basado en dias de cobertura
        if vel_dia > 0:
            if dias_cob < DIAS_CRITICOS:
                semaforo = 'rojo'
            elif dias_cob < DIAS_ALERTA:
                semaforo = 'amarillo'
            else:
                semaforo = 'verde'
        else:
            # Sin ventas: verde si hay stock, amarillo si no hay stock ni ventas
            semaforo = 'verde' if stock_uds > 0 else 'amarillo'

        # Alertas
        if dias_cob < DIAS_CRITICOS and vel_dia > 0:
            dias_str = str(int(dias_cob))
            vel_str = str(int(vel_mes))
            stock_str = str(stock_uds)
            all_alerts.append({
                'producto': prod,
                'nivel': 'critico',
                'tipo': 'cobertura_critica',
                'mensaje': ("Stock: " + stock_str + " uds | " +
                            dias_str + " dias cobertura | Vende " + vel_str + " uds/mes"),
            })

        if missing_mp and can_produce is False:
            n_str = str(len(missing_mp))
            names = ", ".join(m['nombre'] + " -" + str(int(m['deficit_g'])) + "g"
                              for m in missing_mp[:3])
            if len(missing_mp) > 3:
                names += " y " + str(len(missing_mp) - 3) + " mas"
            # Detectar si alguna MP faltante es de China (lead time 60d).
            # Si sí, escalar a CRITICO (no "alto") porque ya estás tarde.
            china_set = china_mps or set()
            mp_china_falta = [m for m in missing_mp
                              if str(m.get('material_id', '')).strip() in china_set]
            if mp_china_falta:
                china_names = ", ".join(m['nombre'] for m in mp_china_falta[:3])
                all_alerts.append({
                    'producto': prod,
                    'nivel': 'critico',
                    'tipo': 'mp_china_faltante',
                    'mensaje': (
                        f"URGENTE — MP de China sin stock: {china_names}. "
                        f"Lead time {LEAD_TIME_CHINA} días. Comprar HOY o se "
                        f"detiene la línea."
                    ),
                })
            all_alerts.append({
                'producto': prod,
                'nivel': 'alto',
                'tipo': 'mp_faltante',
                'mensaje': ("Faltan " + n_str + " MPs para producir: " + names),
            })

        # Alertas anticipadas: para MPs que SÍ tienen stock hoy pero el deficit
        # proyectado va a llegar antes que el lead time del proveedor.
        # Si una MP de China tiene cobertura proyectada < 60d, ya hay que pedir.
        if can_produce is not False and vel_dia > 0:
            china_set = china_mps or set()
            for m in mp_check:
                if not m.get('mp_found', True):
                    continue
                mid = str(m.get('material_id', '')).strip()
                avail = m.get('available_g')
                needed = m.get('needed_g', 0)
                if avail == '∞' or avail == float('inf') or needed <= 0:
                    continue
                # Cuántas producciones futuras puedo cubrir con la MP en stock
                if needed > 0:
                    producciones_cubiertas = float(avail or 0) / needed
                else:
                    continue
                # Días hasta agotamiento de ESA MP (asumiendo ritmo de prod actual)
                # Aproximación: vel_dia uds/día * needed_g/uds = consumo diario MP
                # Pero como needed_g viene del lote, normalizamos: días = avail / (consumo_diario)
                consumo_diario_g = (vel_dia * needed) / max(stock_uds or 1, 1) if stock_uds else 0
                if consumo_diario_g <= 0:
                    continue
                dias_mp = float(avail or 0) / consumo_diario_g
                lead = LEAD_TIME_CHINA if mid in china_set else LEAD_TIME_LOCAL
                if dias_mp < lead:
                    all_alerts.append({
                        'producto': prod,
                        'nivel': 'alto' if mid in china_set else 'medio',
                        'tipo': 'mp_anticipada_china' if mid in china_set else 'mp_anticipada',
                        'mensaje': (
                            f"PEDIR YA: {m['nombre']} dura ~{int(dias_mp)}d, "
                            f"lead time proveedor {lead}d "
                            f"({'China' if mid in china_set else 'local'})."
                        ),
                    })

        if dias_cob < DIAS_ALERTA and vel_dia > 0 and prox_prod == 'No programado':
            vel_str = str(int(vel_mes))
            all_alerts.append({
                'producto': prod,
                'nivel': 'medio',
                'tipo': 'sin_programar',
                'mensaje': ("Vende " + vel_str + " uds/mes y no tiene produccion en calendario"),
            })

        # Proyecciones de unidades faltantes para cubrir N dias
        # vel_dia = unidades vendidas/dia. necesarias_Nd = ceil(vel_dia * N).
        # faltante_Nd = max(0, necesarias_Nd - stock_uds)
        import math as _math
        def _faltante_uds(N):
            if vel_dia <= 0:
                return 0
            necesarias = _math.ceil(vel_dia * N)
            return max(0, necesarias - stock_uds)
        faltante_15d = _faltante_uds(15)
        faltante_30d = _faltante_uds(30)
        faltante_60d = _faltante_uds(60)

        projection.append({
            'producto': prod,
            'lote_kg': lote_kg,
            'vel_mes': round(vel_mes, 1),
            'vel_dia': round(vel_dia, 2),
            'stock_actual': stock_uds,
            'dias_cobertura': round(dias_cob, 0) if dias_cob < 999 else None,
            'faltante_uds_15d': faltante_15d,
            'faltante_uds_30d': faltante_30d,
            'faltante_uds_60d': faltante_60d,
            'prox_produccion': prox_prod,
            'prox_prod_pasada': bool(next_prod_by_product.get(prod + '__past')),
            'cal_ok': cal_ok,
            'mp_lista': can_produce,
            'mp_data_gap': has_data_gap,
            'n_mp_faltantes': len(missing_mp),
            'n_mp_sin_datos': len([m for m in mp_check if not m.get('mp_found', True)]),
            'semaforo': semaforo,
            'mp_check': mp_check,
        })

    order = {'rojo': 0, 'amarillo': 1, 'verde': 2}
    def _sort_key(x):
        sin_fecha = 1 if x['prox_produccion'] == 'No programado' else 0
        return (sin_fecha, order.get(x['semaforo'], 3), x['producto'])
    projection.sort(key=_sort_key)
    all_alerts.sort(key=lambda x: {'critico': 0, 'alto': 1, 'medio': 2}.get(x['nivel'], 3))

    return projection, all_alerts

# ─── Anthropic narrative ─────────────────────────────────────────────────────

# Proveedores China — lead time 60 dias
PROVEEDORES_CHINA = {'lyphar', 'yitibio'}

def _get_china_mps(conn):
    """Retorna set de material_id cuyo proveedor es chino (Lyphar, Yitibio)."""
    china = set()
    try:
        for row in conn.execute(
            "SELECT id, proveedor FROM maestro_mps WHERE proveedor IS NOT NULL"
        ).fetchall():
            prov = str(row[1] or '').lower().strip()
            if any(c in prov for c in PROVEEDORES_CHINA):
                china.add(str(row[0]).strip())
    except Exception:
        pass
    return china


def _generate_narrative(projection, alerts, vel_data, conn=None):
    """
    Analisis IA con horizonte 2 meses.
    Perspectiva: Espagiria recibe demanda de ANIMUS y debe garantizar produccion
    con 20 dias de anticipacion al agotamiento.
    MPs de Lyphar/Yitibio = China = 60 dias lead time.
    """
    if not ANTHROPIC_API_KEY:
        return None

    # Productos criticos (menos de 20 dias)
    criticos = [p for p in projection if p['semaforo'] == 'rojo' and p['vel_dia'] > 0]
    # Productos en alerta (20-40 dias)
    en_alerta = [p for p in projection if p['semaforo'] == 'amarillo' and p['vel_dia'] > 0]
    # MPs faltantes
    mp_faltantes_criticos = [a for a in alerts if a['tipo'] == 'mp_faltante']
    # Sin programar
    sin_programar = [a for a in alerts if a['tipo'] == 'sin_programar']

    # China MPs check
    china_mps = _get_china_mps(conn) if conn else set()
    china_alertas = []
    for p in projection:
        for mp in p.get('mp_check', []):
            if str(mp['material_id']) in china_mps and not mp['ok']:
                china_alertas.append(p['producto'] + ': ' + mp['nombre'] + ' (deficit ' + str(int(mp['deficit_g'])) + 'g de China)')

    # Build data block (no f-strings con chars especiales)
    lines = []
    lines.append('CONTEXTO EMPRESARIAL:')
    lines.append('- ANIMUS Lab: ecommerce que vende por Shopify')
    lines.append('- Espagiria: laboratorio maquila que produce para ANIMUS')
    lines.append('- Regla critica: producir 20 dias ANTES del agotamiento')
    lines.append('- MPs de China (Lyphar, Yitibio): lead time 60 dias — comprar con 2 meses de anticipacion')
    lines.append('')
    lines.append('ESTADO ACTUAL (horizonte 60 dias):')
    lines.append('- Productos criticos (<20 dias stock): ' + str(len(criticos)))
    for p in criticos[:4]:
        dc = str(int(p['dias_cobertura'])) if p['dias_cobertura'] else '?'
        lines.append('  * ' + p['producto'] + ': ' + dc + ' dias, vende ' + str(int(p['vel_mes'])) + ' uds/mes, cal=' + p['prox_produccion'])
    lines.append('- En alerta (20-40 dias): ' + str(len(en_alerta)))
    for p in en_alerta[:3]:
        dc = str(int(p['dias_cobertura'])) if p['dias_cobertura'] else '?'
        lines.append('  * ' + p['producto'] + ': ' + dc + ' dias')
    lines.append('- MPs faltantes para producir: ' + str(len(mp_faltantes_criticos)))
    if mp_faltantes_criticos:
        for a in mp_faltantes_criticos[:3]:
            lines.append('  * ' + a['producto'] + ': ' + a['mensaje'][:80])
    lines.append('- Sin programar en calendario: ' + str(len(sin_programar)))
    lines.append('- MPs de China con deficit: ' + str(len(china_alertas)))
    for ca in china_alertas[:3]:
        lines.append('  * ' + ca)
    lines.append('- Pedidos Shopify ultimos 60d: ' + str(vel_data.get('total_orders', 0)))

    lines.append('')
    lines.append('INSTRUCCION: Genera un analisis ejecutivo en espanol para Espagiria con:')
    lines.append('1. Cuales productos debe producir primero y por que (dias de cobertura)')
    lines.append('2. Que MPs comprar YA (especialmente China — deben pedirse hoy para tenerlos en 60 dias)')
    lines.append('3. Una accion concreta para los proximos 7 dias')
    lines.append('Tono: tecnico-operativo directo. Maximo 5 oraciones.')

    prompt = chr(10).join(lines)

    body = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 400,
        "messages": [{"role": "user", "content": prompt}]
    }).encode('utf-8')

    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=body,
            headers={
                "x-api-key": ANTHROPIC_API_KEY.strip(),
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = json.loads(r.read())
        return resp.get('content', [{}])[0].get('text', '')
    except urllib.error.HTTPError as e:
        body_err = e.read().decode('utf-8', errors='replace')[:200]
        return '[IA error ' + str(e.code) + ': ' + body_err + ']'
    except Exception as e:
        return '[IA no disponible: ' + str(e) + ']' 

# ─── Routes ──────────────────────────────────────────────────────────────────

@bp.route('/api/programacion/resumen')
def prog_resumen():
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401

    conn = get_db()

    # 1. Shopify velocity
    vel_data = _shopify_velocity(conn, days=60)

    # 2. Calendar events
    cal = _fetch_calendar_events(days_ahead=90)

    # 3. MP stock
    mp_stock = _get_mp_stock(conn)

    # 4. Formulas
    formulas = _get_formulas(conn)

    if not formulas:
        return jsonify({
            'error': 'No hay fórmulas cargadas. Las fórmulas se cargan automáticamente al arrancar la app.',
            'formulas_count': 0,
        }), 200

    # 5. Project stock + alerts (pasamos china_mps para alertas anticipadas)
    china_mps_set = _get_china_mps(conn)
    projection, alerts = _project_stock(
        conn, vel_data['prod_velocity'], formulas, mp_stock, cal.get('events', []),
        china_mps=china_mps_set
    )

    # 6. AI narrative (non-blocking)
    narrativa = None
    try:
        narrativa = _generate_narrative(projection, alerts, vel_data, conn=conn)
    except Exception:
        pass

    # Summary KPIs
    n_alertas = len([a for a in alerts if a['nivel'] in ('critico', 'alto')])
    total_vel = sum(vel_data['sku_velocity'].values())
    prox_prod_dates = sorted(
        [p['prox_produccion'] for p in projection if p['prox_produccion'] != 'No programado']
    )
    proxima = prox_prod_dates[0] if prox_prod_dates else 'Sin programar'

    # Warnings de integridad de datos (no son alertas operacionales)
    warnings_data = []
    try:
        collisions = _detect_alias_collisions(conn)
        if collisions:
            warnings_data.append({
                'tipo': 'alias_collision',
                'severidad': 'alta',
                'mensaje': (
                    f"{len(collisions)} grupo(s) de MPs con nombres distintos "
                    f"colapsan al mismo nombre normalizado. El alias map puede "
                    f"hacer que el stock de uno se atribuya al otro silenciosamente."
                ),
                'detalle': collisions[:10],
                'accion': "GET /api/programacion/diagnostico-alias para detalle completo",
            })
    except Exception as _e_w:
        logging.getLogger('programacion').error("warnings_data falló: %s", _e_w)
    if cal.get('error'):
        warnings_data.append({
            'tipo': 'calendar_error',
            'severidad': 'media',
            'mensaje': f"Calendario inaccesible: {cal['error']}",
            'accion': "Las próximas producciones pueden estar incompletas. Verifica GCAL_ICAL_URL o GOOGLE_API_KEY en env.",
        })
    # Warning si la velocidad de ventas viene de datos pobres
    dq = vel_data.get('data_quality', 'ok')
    if dq != 'ok':
        cov = vel_data.get('coverage_pct', 0)
        warnings_data.append({
            'tipo': 'velocidad_data_pobre',
            'severidad': 'alta' if dq in ('no_data', 'very_low') else 'media',
            'mensaje': (
                f"Velocidad de ventas calculada con datos insuficientes "
                f"(calidad={dq}, cobertura={cov}%). Las decisiones de "
                f"producción/compra basadas en esto pueden estar sesgadas."
            ),
            'accion': "Verifica el sync con Shopify (/api/programacion/test-shopify).",
        })
    # Fórmulas con cantidades incompletas
    formulas_incompletas = [
        p for p in projection
        if p.get('mp_lista') is None  # can_produce=None → fórmula sin cantidades
    ]
    if formulas_incompletas:
        warnings_data.append({
            'tipo': 'formulas_incompletas',
            'severidad': 'alta',
            'mensaje': (
                f"{len(formulas_incompletas)} fórmula(s) tienen cantidades vacías "
                f"o inválidas. La proyección de MPs para esos productos no es "
                f"confiable."
            ),
            'productos': [p['producto'] for p in formulas_incompletas[:10]],
            'accion': "Edita la fórmula en /tecnica y completa las cantidades por lote.",
        })

    return jsonify({
        'velocidad_total': round(total_vel, 0),
        'proxima_produccion': proxima,
        'n_alertas': n_alertas,
        'formulas_count': len(formulas),
        'mp_stock_count': len(mp_stock),
        'proyeccion': projection,
        'alertas': alerts,
        'velocidad': vel_data,
        'calendario': cal,
        'calendario_ok': cal.get('error') is None,
        'narrativa_ia': narrativa,
        'warnings_datos': warnings_data,
        'thresholds': {
            'dias_criticos': DIAS_CRITICOS,
            'dias_alerta': DIAS_ALERTA,
            'lead_time_china': LEAD_TIME_CHINA,
            'lead_time_local': LEAD_TIME_LOCAL,
        },
    })


@bp.route('/api/programacion/diagnostico-alias')
def prog_diagnostico_alias():
    """Lista colisiones del alias map: 2+ MPs con nombres distintos que
    normalizan al mismo nombre. Si aparecen, el stock de uno puede
    atribuirse silenciosamente al otro.
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()
    collisions = _detect_alias_collisions(conn)
    return jsonify({
        'collisions': collisions,
        'count': len(collisions),
        'accion_sugerida': (
            "Para cada grupo, verifica que sean realmente la misma MP. Si lo son, "
            "estandariza el nombre en maestro_mps. Si NO son la misma, renombra "
            "una para que el normalizador no las colapse."
        ),
    })




@bp.route('/api/programacion/que-puedo-producir', methods=['GET'])
def prog_que_puedo_producir():
    """Para cada producto con formula, evalua si las MPs alcanzan para
    producir 1 lote estandar y devuelve faltantes detallados + shopping
    list por proveedor.

    Verificacion paso a paso (auditable):
      1. SOURCE: SELECT material_id, SUM(...) FROM movimientos
         (mismo query que el dashboard usa — kardex actual)
      2. FORMULA: SELECT FROM formula_items WHERE producto=?
         (los gramos requeridos por lote vienen de cantidad_g_por_lote)
      3. MATCH: para cada MP requerida, busca su stock por codigo_mp
         O por nombre (fallback) — incluye los lotes individuales como
         evidencia
      4. PROVEEDOR: cae al canonico de maestro_mps

    Cada MP en la respuesta trae:
      - requerido_g (de formula)
      - stock_actual_g (de kardex)
      - lotes_disponibles[]: evidencia auditable
      - falta_g (negativo => sobra; positivo => falta esa cantidad)
      - ok: true si stock >= requerido

    Producto.puede_producir = todas las MPs ok.
    Si falla, shopping_list agrupa por proveedor lo que hay que comprar.

    Query params:
      cantidad_kg_override: int (opcional) — simular un lote distinto
                            al estandar de la formula
      solo_faltantes: 1 — devuelve solo productos que NO se pueden producir
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401

    conn = get_db()
    c = conn.cursor()

    cantidad_kg_override = request.args.get('cantidad_kg_override', '').strip()
    try:
        cantidad_kg_override = float(cantidad_kg_override) if cantidad_kg_override else None
    except ValueError:
        cantidad_kg_override = None
    solo_faltantes = request.args.get('solo_faltantes', '0') in ('1', 'true', 'True')

    # Step 1: stock por material_id (canonical)
    mp_stock = _get_mp_stock(conn)

    # Step 2: lotes individuales por material_id (evidencia auditable)
    lotes_por_mp = {}
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,''),
                   SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as neto,
                   MAX(fecha_vencimiento), MAX(proveedor)
            FROM movimientos
            WHERE material_id IS NOT NULL AND material_id != ''
            GROUP BY material_id, lote
            HAVING neto > 0
            ORDER BY material_id, fecha_vencimiento ASC
        """).fetchall()
        for r in rows:
            mid = (r[0] or '').strip()
            lotes_por_mp.setdefault(mid, []).append({
                'lote': r[1] or '',
                'cantidad_g': round(float(r[2] or 0), 1),
                'fecha_venc': str(r[3])[:10] if r[3] else '',
                'proveedor': r[4] or '',
            })
    except sqlite3.OperationalError:
        pass

    # Step 3: catalogo proveedor canonico
    cat_proveedor = {}
    try:
        for r in c.execute(
            "SELECT codigo_mp, COALESCE(proveedor,''), COALESCE(nombre_comercial,'') "
            "FROM maestro_mps"
        ).fetchall():
            cat_proveedor[r[0]] = {'proveedor': r[1], 'nombre': r[2]}
    except sqlite3.OperationalError:
        pass

    # Step 4: formulas
    formulas = _get_formulas(conn)

    productos = []
    for nombre_prod, info in formulas.items():
        items = info.get('items', [])
        if not items:
            continue
        lote_size_kg = info.get('lote_size_kg') or 0
        cantidad_kg = cantidad_kg_override if cantidad_kg_override else lote_size_kg
        if cantidad_kg <= 0:
            continue

        factor = cantidad_kg / lote_size_kg if lote_size_kg > 0 else 1
        mps_status = []
        falta_total_g = 0.0
        n_ok = 0
        n_falta = 0

        for it in items:
            mid = (it.get('material_id') or '').strip()
            req_g_estandar = float(it.get('cantidad_g_por_lote') or 0)
            req_g = round(req_g_estandar * factor, 2)

            # Stock por id; fallback a nombre si no existe id
            stock_g = mp_stock.get(mid, 0.0)
            if not mid:
                # Buscar por nombre canonical via mp_stock dict (which also indexes by name)
                stock_g = mp_stock.get((it.get('material_nombre') or '').upper().strip(), 0.0)

            falta_g = round(req_g - stock_g, 2)
            ok = stock_g >= req_g

            if ok:
                n_ok += 1
            else:
                n_falta += 1
                if falta_g > 0:
                    falta_total_g += falta_g

            cat_info = cat_proveedor.get(mid, {})
            mps_status.append({
                'codigo_mp': mid,
                'nombre': cat_info.get('nombre') or it.get('material_nombre') or mid,
                'requerido_g': req_g,
                'stock_actual_g': round(stock_g, 1),
                'falta_g': falta_g if falta_g > 0 else 0,
                'sobra_g': abs(falta_g) if falta_g < 0 else 0,
                'ok': ok,
                'lotes_disponibles': lotes_por_mp.get(mid, [])[:5],
                'proveedor_canonico': cat_info.get('proveedor', ''),
            })

        puede_producir = (n_falta == 0)

        if solo_faltantes and puede_producir:
            continue

        productos.append({
            'producto': nombre_prod,
            'lote_size_kg': lote_size_kg,
            'cantidad_kg_evaluada': cantidad_kg,
            'puede_producir': puede_producir,
            'mps_total': len(items),
            'mps_ok': n_ok,
            'mps_faltantes': n_falta,
            'falta_total_g': round(falta_total_g, 1),
            'mps_status': sorted(mps_status, key=lambda m: (m['ok'], -m['falta_g'])),
        })

    # Step 5: shopping list por proveedor
    shopping = {}
    for prod in productos:
        for m in prod['mps_status']:
            if not m['ok'] and m['falta_g'] > 0:
                prov = m['proveedor_canonico'] or '(sin proveedor)'
                bucket = shopping.setdefault(prov, {})
                k = m['codigo_mp']
                if k in bucket:
                    bucket[k]['falta_g'] = max(bucket[k]['falta_g'], m['falta_g'])
                    bucket[k]['productos_afectados'].add(prod['producto'])
                else:
                    bucket[k] = {
                        'codigo_mp': k, 'nombre': m['nombre'],
                        'falta_g': m['falta_g'],
                        'productos_afectados': {prod['producto']},
                    }
    shopping_list = []
    for prov, mps in sorted(shopping.items()):
        items_list = []
        total_g = 0.0
        for k, item in mps.items():
            items_list.append({
                'codigo_mp': item['codigo_mp'],
                'nombre': item['nombre'],
                'falta_g': round(item['falta_g'], 1),
                'productos_afectados': sorted(item['productos_afectados']),
            })
            total_g += item['falta_g']
        items_list.sort(key=lambda x: -x['falta_g'])
        shopping_list.append({
            'proveedor': prov,
            'count_mps': len(items_list),
            'total_g_a_pedir': round(total_g, 1),
            'mps': items_list,
        })
    shopping_list.sort(key=lambda s: -s['total_g_a_pedir'])

    productos.sort(key=lambda p: (p['puede_producir'], -p['falta_total_g']))

    # KPIs resumen
    n_total = len(productos)
    n_pueden = sum(1 for p in productos if p['puede_producir'])
    n_no_pueden = n_total - n_pueden

    return jsonify({
        'ok': True,
        'fecha': __import__('datetime').date.today().isoformat(),
        'fuente_datos': {
            'kardex': '/api/movimientos (SUM signed por material_id)',
            'formulas': 'formula_items.cantidad_g_por_lote escalado a cantidad_kg',
            'proveedor_canonico': 'maestro_mps.proveedor',
        },
        'resumen': {
            'productos_totales': n_total,
            'productos_pueden_producir': n_pueden,
            'productos_con_faltantes': n_no_pueden,
            'cantidad_kg_override_usada': cantidad_kg_override,
        },
        'productos': productos,
        'shopping_list_por_proveedor': shopping_list,
    })


@bp.route('/api/programacion/registrar-stock', methods=['POST'])
def prog_registrar_stock():
    """Registra stock inicial de PT directamente en stock_pt."""
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    d = request.get_json(force=True) or {}
    sku      = str(d.get('sku', '') or '').strip().upper()
    producto = str(d.get('producto', '') or '').strip()
    unidades = int(d.get('unidades', 0) or 0)
    lote     = str(d.get('lote', 'INI-' + datetime.now().strftime('%Y%m%d')) or '')

    if not sku or not producto or unidades <= 0:
        return jsonify({'error': 'sku, producto y unidades requeridos (unidades > 0)'}), 400

    conn = get_db()
    # Invalidate previous carga-inicial entries for same sku to avoid duplicates
    conn.execute(
        "UPDATE stock_pt SET estado='Ajustado' WHERE sku=? AND lote_produccion LIKE 'INI-%'",
        (sku,)
    )
    conn.execute("""
        INSERT INTO stock_pt
            (sku, descripcion, lote_produccion, fecha_produccion,
             unidades_inicial, unidades_disponible, precio_base,
             empresa, estado, observaciones)
        VALUES (?,?,?,date('now', '-5 hours'),?,?,0,'ANIMUS','Disponible','Carga inicial de stock')
    """, (sku, producto, lote, unidades, unidades))
    conn.commit()
    return jsonify({'ok': True, 'sku': sku, 'producto': producto, 'unidades': unidades})



def _fetch_shopify_available(token, shop, inv_item_ids, location_id=None):
    """Para una lista de inventory_item_ids de Shopify, devuelve dict
    {inv_item_id: available}.

    Sebastián 12-may-2026: fix #D On hand → Available. Shopify Admin API
    expone 'inventory_quantity' = ON HAND (incluye Committed/Reservado),
    pero para MRP/planeación necesitamos AVAILABLE (= On hand - Committed).
    Para obtenerlo hay que ir a /inventory_levels.json que sí trae 'available'.

    FIX 1-jun-2026 (Sebastián · caso LBHA): si NO se pasa location_id, sumaba
    'available' de TODAS las locations. Con una location fantasma/vieja en negativo
    el total quedaba absurdo (ej. LBHA: location real 226 + fantasma -461 = -235) y
    el motor mostraba stock 0/erróneo. Ahora, si se pasa location_id, se filtra a
    esa única location (la real de la tienda) vía &location_ids=. Sin location_id
    mantiene el comportamiento anterior (suma todas) para compatibilidad.

    Limitación API: ~50 IDs por request. Chunkeamos.
    Si falla (red/auth/empty IDs), devuelve dict vacío y loggea warning.
    El caller debe tener fallback (usualmente: usar inventory_quantity).
    """
    out = {}
    if not inv_item_ids:
        return out
    # Dedupe + filtrar None/0
    unique_ids = sorted({int(x) for x in inv_item_ids if x})
    chunk = 50
    _loc_q = ('&location_ids=' + str(location_id)) if location_id else ''
    # FIX 23-may-2026 · auditoría P2 · antes si un chunk fallaba seguía con
    # continue · resultado: mezcla Available + On Hand para distintos SKUs
    # en el mismo sync · ahora all-or-nothing + retry con backoff en 429
    import time as _time
    for i in range(0, len(unique_ids), chunk):
        sub = unique_ids[i:i + chunk]
        ids_csv = ','.join(str(x) for x in sub)
        url = ('https://' + shop +
               '/admin/api/2024-01/inventory_levels.json'
               '?inventory_item_ids=' + ids_csv + _loc_q + '&limit=' + str(chunk * 5))
        # Retry con backoff exponencial para 429/5xx · 3 intentos
        data = None
        for intento in range(3):
            req = urllib.request.Request(url, headers={'X-Shopify-Access-Token': token})
            try:
                with urllib.request.urlopen(req, timeout=20) as r:
                    data = json.loads(r.read())
                break
            except urllib.error.HTTPError as he:
                if he.code in (429, 500, 502, 503, 504) and intento < 2:
                    # Respetar Retry-After si vino · sino backoff exponencial 2/4s
                    ra = 0
                    try:
                        ra = int(he.headers.get('Retry-After', '0') or 0)
                    except Exception:
                        ra = 0
                    _time.sleep(max(ra, 2 ** (intento + 1)))
                    continue
                log.warning("inventory_levels chunk %d-%d HTTP %s: %s",
                            i, i + chunk, he.code, he)
                return {}   # all-or-nothing · caller fallback a On Hand
            except Exception as e:
                if intento < 2:
                    _time.sleep(2 ** (intento + 1))
                    continue
                log.warning("inventory_levels chunk %d-%d falló: %s", i, i + chunk, e)
                return {}   # all-or-nothing
        if data is None:
            return {}
        for lvl in data.get('inventory_levels', []) or []:
            iid = lvl.get('inventory_item_id')
            av = lvl.get('available')
            if iid is None or av is None:
                continue
            try:
                _iid = int(iid)
                _v = int(av)
            except Exception:
                continue
            if location_id:
                # Filtrado a la(s) location(s) reales → sumar (normalmente 1 sola)
                out[_iid] = out.get(_iid, 0) + _v
            else:
                # FIX 1-jun-2026 · sin filtro de location: tomar el MÁXIMO por ítem
                # entre locations en vez de SUMAR. Así una location fantasma/vieja en
                # negativo (caso LBHA: 226 real vs -461 fantasma) ya NO arrastra el
                # total a negativo · la location real (positiva) domina. Robusto aún
                # si el token no tiene scope read_locations para autodetectar.
                if _iid not in out or _v > out[_iid]:
                    out[_iid] = _v
    return out


def _shopify_locations(token, shop, timeout=12):
    """Lista de locations de Shopify: [{'id','name','active','legacy'}]. [] si falla."""
    try:
        url = 'https://' + shop + '/admin/api/2024-01/locations.json'
        req = urllib.request.Request(url, headers={'X-Shopify-Access-Token': token})
        with urllib.request.urlopen(req, timeout=timeout) as r:
            data = json.loads(r.read())
        return [{'id': l.get('id'), 'name': l.get('name', ''),
                 'active': bool(l.get('active')), 'legacy': bool(l.get('legacy'))}
                for l in (data.get('locations') or []) if l.get('id')]
    except Exception:
        return []


def _shopify_location_id(conn, token, shop):
    """location_id de la tienda ÁNIMUS LAB para leer Available · Sebastián 1-jun-2026:
    'solo debe mirar tienda animus lab nada más'. Antes se sumaba el Available de TODAS
    las locations → una location fantasma/vieja en negativo daba totales absurdos
    (LBHA: 226 real + -461 fantasma = -235). Prioriza animus_config('shopify_location_id');
    si no, autodetecta la location cuyo nombre contiene 'ANIMUS'. Devuelve str o None."""
    try:
        r = conn.execute(
            "SELECT valor FROM animus_config WHERE clave='shopify_location_id'").fetchone()
        if r and (r[0] or '').strip():
            return (r[0] or '').strip()
    except Exception:
        pass

    def _norm(s):
        import unicodedata
        return ''.join(ch for ch in unicodedata.normalize('NFKD', str(s or ''))
                       if not unicodedata.combining(ch)).upper().strip()

    locs = _shopify_locations(token, shop)
    for l in locs:
        if 'ANIMUS' in _norm(l.get('name')) and l.get('id'):
            return str(l['id'])
    activas = [str(l['id']) for l in locs
               if l.get('active') and not l.get('legacy') and l.get('id')]
    if len(activas) == 1:
        return activas[0]
    return None


def _stamp_stock_sync(conn):
    """Marca el momento (UTC) del último sync de stock Shopify · habilita el
    auto-refresh en vivo de Necesidades."""
    try:
        from datetime import datetime as _dt
        conn.execute(
            "INSERT OR REPLACE INTO animus_config (clave, valor) VALUES ('last_stock_sync_at', ?)",
            (_dt.utcnow().isoformat(timespec='seconds'),))
    except Exception:
        pass


def _auto_refresh_shopify_stock(conn, max_age_seg=600):
    """Necesidades EN VIVO (Sebastián 1-jun-2026: 'que lea Shopify en vivo, no el
    snapshot'). Si el snapshot de stock Shopify (stock_pt) está más viejo que
    max_age_seg, lo refresca llamando al sync real (con el fix de location/máximo).
    · Best-effort: NUNCA rompe ni bloquea Necesidades si Shopify falla.
    · Lock-guarded: SOLO una carga/worker sincroniza · las demás usan el snapshot
      actual sin esperar (no se cuelga la vista para todos)."""
    try:
        from datetime import datetime as _dt
        r = conn.execute(
            "SELECT valor FROM animus_config WHERE clave='last_stock_sync_at'").fetchone()
        if r and r[0]:
            try:
                if (_dt.utcnow() - _dt.fromisoformat(r[0])).total_seconds() < max_age_seg:
                    return  # snapshot fresco · no re-sincronizar
            except Exception:
                pass
        try:
            from blueprints.auto_plan_jobs import (
                _adquirir_lock_cron, _liberar_lock_cron, job_sync_stock_shopify_diario)
        except Exception:
            from api.blueprints.auto_plan_jobs import (  # type: ignore
                _adquirir_lock_cron, _liberar_lock_cron, job_sync_stock_shopify_diario)
        if not _adquirir_lock_cron(conn, 'sync_stock_shopify', ttl_horas=1):
            return  # otra carga/worker ya está sincronizando · usar snapshot actual
        try:
            from flask import current_app
            job_sync_stock_shopify_diario(current_app._get_current_object())
        finally:
            try:
                _liberar_lock_cron(conn, 'sync_stock_shopify')
            except Exception:
                pass
    except Exception:
        pass


@bp.route('/api/programacion/test-shopify')
def prog_test_shopify():
    """GET diagnostico: verifica credenciales y cuenta productos Shopify.

    SEC-FIX 23-may-2026 · auditoría · era público · exponía shop name +
    token prefix (8 chars) · reconocimiento de credenciales.
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    try:
        conn = get_db()
        def _cfg(c):
            r = conn.execute("SELECT valor FROM animus_config WHERE clave=?", (c,)).fetchone()
            return r[0] if r else None
        token = _cfg('shopify_token')
        shop  = _cfg('shopify_shop')
        if not token or not shop:
            return jsonify({'ok': False, 'paso': 'credenciales', 'error': 'shopify_token o shopify_shop no configurados en animus_config'})
        url = 'https://' + shop + '/admin/api/2024-01/products/count.json'
        req = urllib.request.Request(url, headers={'X-Shopify-Access-Token': token})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())
        return jsonify({'ok': True, 'shop': shop, 'count': data.get('count', '?'), 'token_prefix': token[:8] + '...'})
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')[:400]
        return jsonify({'ok': False, 'paso': 'shopify_api', 'http_status': e.code, 'body': body})
    except Exception as e:
        return jsonify({'ok': False, 'paso': 'exception', 'error': str(e)})


@bp.route('/api/programacion/sync-salud', methods=['GET'])
def prog_sync_salud():
    """Salud del sync Shopify (local, sin llamada externa) + diagnóstico del filtro
    B2B (SHOPIFY_B2B_TAGS): qué tags traen realmente las órdenes, cuántas se
    clasifican B2B y si la var está configurada. Sebastián 31-may-2026."""
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    import os as _os
    conn = get_db(); c = conn.cursor()
    out = {'ok': True}
    try:
        def _cfg(k):
            r = conn.execute("SELECT valor FROM animus_config WHERE clave=?", (k,)).fetchone()
            return r[0] if r else None
        out['config'] = {'dominio_set': bool(_cfg('shopify_shop')),
                         'token_set': bool(_cfg('shopify_token'))}
    except Exception as e:
        out['config'] = {'error': str(e)}
    try:
        out['ordenes_total'] = c.execute("SELECT COUNT(*) FROM animus_shopify_orders").fetchone()[0]
    except Exception:
        out['ordenes_total'] = None
    try:
        out['ordenes_30d'] = c.execute(
            "SELECT COUNT(*) FROM animus_shopify_orders "
            "WHERE creado_en >= date('now','-30 days')").fetchone()[0]
    except Exception:
        out['ordenes_30d'] = None
    try:
        out['ultima_sync'] = c.execute(
            "SELECT MAX(synced_at) FROM animus_shopify_orders").fetchone()[0]
    except Exception:
        out['ultima_sync'] = None
    try:
        b2b_raw = (_os.environ.get('SHOPIFY_B2B_TAGS') or '').strip()
        tags_cfg = [t.strip().lower() for t in b2b_raw.split(',') if t.strip()]
        b2b = {'tags_configurados': tags_cfg, 'configurado': bool(tags_cfg)}
        cols_ok = True
        try:
            c.execute("SELECT tags, customer_tags FROM animus_shopify_orders LIMIT 1")
        except Exception:
            cols_ok = False
        b2b['columnas_tags_existen'] = cols_ok
        if cols_ok:
            from collections import Counter as _Counter
            rows = c.execute(
                "SELECT COALESCE(tags,''), COALESCE(customer_tags,'') "
                "FROM animus_shopify_orders WHERE creado_en >= date('now','-30 days')"
            ).fetchall()
            con_tag = 0; clasif_b2b = 0; vistos = _Counter()
            for tg, ctg in rows:
                full = (str(tg) + ',' + str(ctg)).lower()
                indiv = [x.strip() for x in full.split(',') if x.strip()]
                if indiv:
                    con_tag += 1
                    for x in set(indiv):
                        vistos[x] += 1
                if tags_cfg and any(cfg in full for cfg in tags_cfg):
                    clasif_b2b += 1
            b2b['ordenes_30d_con_tag'] = con_tag
            b2b['ordenes_30d_clasificadas_b2b'] = clasif_b2b
            b2b['tags_vistos_top'] = vistos.most_common(20)
        out['b2b'] = b2b
    except Exception as e:
        out['b2b'] = {'error': str(e)}

    # ── Salud del SYNC DE STOCK (stock_pt SHOPIFY) ──────────────────────
    # Sebastián 1-jun-2026 · "necesito que jale los stock reales · día a día".
    # Revela si el sync de stock corrió, si usó Available o cayó a On hand, y
    # si TODO quedó Agotado (señal típica de token sin scope read_inventory).
    try:
        stk = {}
        stk['ultimo_sync_stock'] = c.execute(
            "SELECT MAX(fecha_produccion) FROM stock_pt "
            "WHERE lote_produccion LIKE 'SHOPIFY-%'").fetchone()[0]
        stk['skus_disponibles'] = c.execute(
            "SELECT COUNT(DISTINCT UPPER(TRIM(sku))) FROM stock_pt "
            "WHERE lote_produccion LIKE 'SHOPIFY-%' AND estado='Disponible'").fetchone()[0]
        stk['skus_agotados'] = c.execute(
            "SELECT COUNT(DISTINCT UPPER(TRIM(sku))) FROM stock_pt "
            "WHERE lote_produccion LIKE 'SHOPIFY-%' AND estado='Agotado'").fetchone()[0]
        stk['uds_disponibles_total'] = c.execute(
            "SELECT COALESCE(SUM(unidades_disponible),0) FROM stock_pt "
            "WHERE lote_produccion LIKE 'SHOPIFY-%' AND estado='Disponible'").fetchone()[0]
        fuente = c.execute(
            "SELECT COALESCE(observaciones,''), COUNT(*) FROM stock_pt "
            "WHERE lote_produccion LIKE 'SHOPIFY-%' AND estado IN ('Disponible','Agotado') "
            "GROUP BY COALESCE(observaciones,'') ORDER BY COUNT(*) DESC LIMIT 4").fetchall()
        stk['fuente_observaciones'] = [{'obs': o, 'n': n} for o, n in fuente]
        stk['uso_available'] = any('Available' in (o or '') for o, _ in fuente)
        if (stk['skus_agotados'] or 0) > 0 and (stk['skus_disponibles'] or 0) == 0:
            stk['alerta'] = ('TODOS los SKU quedaron Agotados (0 disponibles). Causa típica: '
                             'el token Shopify no tiene scope read_inventory → Available/On hand '
                             'vuelve 0. Revisá shopify_scopes abajo.')
        out['stock'] = stk
    except Exception as e:
        out['stock'] = {'error': str(e)}

    # ── Scopes del token (llamada externa liviana) · confirma read_inventory ──
    # Es el chequeo decisivo: sin read_inventory, Shopify no devuelve
    # Available ni On hand → todo el stock entra en 0.
    try:
        token = _cfg('shopify_token')
        shop = _cfg('shopify_shop')
        if token and shop:
            url = 'https://' + shop + '/admin/oauth/access_scopes.json'
            req = urllib.request.Request(url, headers={'X-Shopify-Access-Token': token})
            with urllib.request.urlopen(req, timeout=12) as r:
                sc = json.loads(r.read())
            handles = [s.get('handle') for s in sc.get('access_scopes', []) if s.get('handle')]
            sc_out = {
                'todos': sorted(handles),
                'read_products': 'read_products' in handles,
                'read_inventory': 'read_inventory' in handles,
            }
            if 'read_inventory' not in handles:
                sc_out['alerta'] = ('⚠ FALTA read_inventory · sin él Shopify NO devuelve '
                    'Available/On hand → el stock de Necesidades queda en 0. Agregá el scope '
                    'en la Custom App de Shopify (Apps → tu app → API scopes) y reinstalá/'
                    'regenerá el token.')
            out['shopify_scopes'] = sc_out
        else:
            out['shopify_scopes'] = {'error': 'token o shop no configurados'}
    except urllib.error.HTTPError as e:
        out['shopify_scopes'] = {'error': 'HTTP ' + str(e.code) + ' al leer access_scopes (token inválido o sin permiso)'}
    except Exception as e:
        out['shopify_scopes'] = {'error': str(e)}
    return jsonify(out)


@bp.route('/api/programacion/reconciliar-shopify', methods=['GET'])
def prog_reconciliar_shopify():
    """Read-only · trae EN VIVO cada variante de Shopify y la reconcilia contra lo
    que el motor de Necesidades ve. Por SKU: On hand, Available, si está mapeada,
    a qué producto, qué stock resuelve el motor, cuánto vende (60d) y un diagnóstico.
    Sebastián 1-jun-2026: 'revisemos cómo jala de Shopify todo y cada uno'."""
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()
    q_filtro = (request.args.get('q') or '').strip().upper()

    def _cfg(k):
        r = conn.execute("SELECT valor FROM animus_config WHERE clave=?", (k,)).fetchone()
        return r[0] if r else None

    token = _cfg('shopify_token')
    shop = _cfg('shopify_shop')
    if not token or not shop:
        return jsonify({'ok': False, 'error': 'Shopify no configurado'}), 200

    # Mapeos locales
    sku_map = {}
    for row in conn.execute(
            "SELECT sku, producto_nombre FROM sku_producto_map WHERE COALESCE(activo,1)=1").fetchall():
        sku_map[str(row[0] or '').strip().upper()] = str(row[1] or '').strip()
    pres_map = {}
    try:
        for row in conn.execute(
                "SELECT UPPER(TRIM(sku_shopify)), producto_nombre FROM producto_presentaciones "
                "WHERE sku_shopify IS NOT NULL AND TRIM(sku_shopify)!=''").fetchall():
            if row[0]:
                pres_map[row[0]] = row[1]
    except Exception:
        pass

    # Variantes en vivo desde Shopify
    variants = []
    url = ('https://' + shop +
           '/admin/api/2024-01/products.json?limit=250&fields=id,title,variants')
    try:
        while url:
            req = urllib.request.Request(url, headers={'X-Shopify-Access-Token': token})
            with urllib.request.urlopen(req, timeout=20) as r:
                data = json.loads(r.read())
                link = r.headers.get('Link', '') or ''
            for p in data.get('products', []):
                title = str(p.get('title', '') or '')
                for v in p.get('variants', []):
                    variants.append({
                        'sku': str(v.get('sku', '') or '').strip().upper(),
                        'title': title,
                        'on_hand': int(v.get('inventory_quantity', 0) or 0),
                        'inv_item_id': v.get('inventory_item_id'),
                        'inv_mgmt': v.get('inventory_management'),
                    })
            nxt = None
            for part in link.split(','):
                if 'rel="next"' in part:
                    s = part.find('<') + 1
                    e = part.find('>')
                    if s > 0 and e > s:
                        nxt = part[s:e].strip()
            url = nxt
    except urllib.error.HTTPError as e:
        return jsonify({'ok': False, 'error': 'Shopify HTTP ' + str(e.code) + ' (revisá scopes/token)'}), 200
    except Exception as e:
        return jsonify({'ok': False, 'error': 'Error red Shopify: ' + str(e)}), 200

    # Available real (inventory_levels) + stock resuelto por el motor
    iids = [v['inv_item_id'] for v in variants if v.get('inv_item_id')]
    loc_id = _shopify_location_id(conn, token, shop)  # solo tienda ÁNIMUS LAB
    avail = _fetch_shopify_available(token, shop, iids, location_id=loc_id)
    used_available = bool(avail)
    resolved = _resolved_stock_por_sku(conn, empresa='ANIMUS')

    # Ventas 60d por SKU (para ver 'vende pero stock 0')
    ventas60 = {}
    try:
        for (si,) in conn.execute(
                "SELECT sku_items FROM animus_shopify_orders "
                "WHERE creado_en >= date('now','-60 days') AND sku_items IS NOT NULL").fetchall():
            try:
                items = json.loads(si)
            except Exception:
                continue
            for it in (items or []):
                k = str(it.get('sku', '') or '').strip().upper()
                q = int(it.get('qty') or it.get('quantity') or 0)
                if k and q > 0:
                    ventas60[k] = ventas60.get(k, 0) + q
    except Exception:
        pass

    filas = []
    n_map = n_sinmap = n_avail_dif = n_problema = 0
    for v in variants:
        sku = v['sku']
        iid = v.get('inv_item_id')
        av = avail.get(int(iid)) if (iid and int(iid) in avail) else None
        prod = sku_map.get(sku) or pres_map.get(sku)
        mapeado = sku in sku_map
        res = resolved.get(sku, {})
        res_uds = int(res.get('uds', 0) or 0)
        vende = ventas60.get(sku, 0)
        shop_qty = av if av is not None else v['on_hand']
        if mapeado:
            n_map += 1
        else:
            n_sinmap += 1
        if av is not None and av != v['on_hand']:
            n_avail_dif += 1
        problema = False
        if not sku:
            diag = 'SKU vacío en Shopify · no se puede mapear'
            problema = True
        elif not prod:
            diag = '⚠ SKU NO mapeado (ni sku_producto_map ni presentaciones) → su stock NO aparece en Necesidades'
            problema = True
        elif res_uds > 0:
            diag = 'OK · motor ve ' + str(res_uds) + ' (' + str(res.get('fuente', '')) + ')'
        elif (shop_qty or 0) <= 0:
            diag = 'Available=0 en Shopify → 0 correcto (toca producir)'
        else:
            diag = ('⚠ Shopify tiene ' + str(shop_qty) + ' pero el motor resuelve 0 · '
                    'la fila en stock_pt no coincide por SKU o no quedó Disponible')
            problema = True
        if problema:
            n_problema += 1
        filas.append({
            'sku': sku, 'producto': prod or '', 'mapeado': mapeado,
            'en_presentaciones': sku in pres_map,
            'on_hand': v['on_hand'], 'available': av,
            'rastrea_inventario': bool(v.get('inv_mgmt')),
            'resuelto_motor': res_uds, 'fuente': res.get('fuente', ''),
            'vende_60d': vende, 'problema': problema, 'diagnostico': diag,
        })
    # problemas primero, luego por producto
    filas.sort(key=lambda x: (not x['problema'], x['producto'], x['sku']))
    # Filtro opcional ?q= · para inspeccionar un producto/SKU puntual (ej. ?q=BHA)
    if q_filtro:
        filas = [f for f in filas
                 if q_filtro in (f.get('sku') or '').upper()
                 or q_filtro in (f.get('producto') or '').upper()]
    return jsonify({
        'ok': True,
        'used_available': used_available,
        'location_id': loc_id,
        'total_variantes': len(variants),
        'mapeados': n_map, 'sin_mapeo': n_sinmap,
        'available_difiere_onhand': n_avail_dif,
        'con_problema': n_problema,
        'onhand_total': sum(v['on_hand'] for v in variants),
        'available_total': (sum(int(x) for x in avail.values()) if avail else None),
        'filas': filas,
    })


@bp.route('/api/programacion/sync-ventas', methods=['POST'])
def prog_sync_ventas():
    """Sincroniza ordenes Shopify directamente — independiente de marketing."""
    if 'compras_user' not in session:
        return jsonify({'ok': False, 'error': 'No autorizado'}), 401
    try:
        days = int(request.json.get('days', 60)) if request.json else 60
        conn = get_db()
        result = _sync_shopify_orders(conn, days=days)
        if result['ok']:
            return jsonify({
                'ok': True,
                'synced': result['synced'],
                'mensaje': str(result['synced']) + ' ordenes sincronizadas (' + str(days) + 'd)'
            })
        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()[-500:]})

@bp.route('/api/programacion/sync-stock-shopify', methods=['POST'])
def prog_sync_stock_shopify():
    """
    Sincroniza inventario desde Shopify products API a stock_pt.
    Acepta sesión 'compras_user' o ?clave=AUTO_PLAN_CRON_KEY (cron Render).
    Siempre retorna JSON; nunca lanza 500 HTML.

    Sebastián 1-may-2026 audit: antes era public · ahora requiere uno de los 2.
    """
    from blueprints.auto_plan import _validar_acceso_cron
    es_cron, _err = _validar_acceso_cron(request)
    if 'compras_user' not in session and not es_cron:
        return jsonify({'ok': False, 'error': 'No autorizado'}), 401
    # FIX 23-may-2026 · auditoría P2 · antes UPDATE+INSERT sin transacción/lock
    # 2 syncs simultáneos (cron + botón manual o 2 admins) abrían ventana donde
    # todo stock quedaba en estado='Ajustado' · _resolved_stock_por_sku devolvía
    # 0 · cobertura de Necesidades se rompía momentáneamente · ahora con
    # cron_locks para serializar
    from blueprints.auto_plan_jobs import _adquirir_lock_cron, _liberar_lock_cron
    _lock_conn = get_db()
    if not _adquirir_lock_cron(_lock_conn, 'sync_stock_shopify', ttl_horas=1):
        return jsonify({
            'ok': False,
            'error': 'Otro sync de stock Shopify está en curso · intenta en unos segundos',
            'reintentar': True,
        }), 409
    try:
        conn = get_db()

        def _cfg(c):
            r = conn.execute("SELECT valor FROM animus_config WHERE clave=?", (c,)).fetchone()
            return r[0] if r else None

        token = _cfg('shopify_token')
        shop  = _cfg('shopify_shop')
        if not token or not shop:
            return jsonify({'ok': False, 'error': 'Shopify no configurado. Ve a ANIMUS > Configuracion y guarda shopify_token y shopify_shop.'})

        sku_map = {}
        for row in conn.execute("SELECT sku, producto_nombre FROM sku_producto_map WHERE activo=1").fetchall():
            sku_map[str(row[0] or '').strip().upper()] = str(row[1] or '').strip()

        all_variants = []
        url = 'https://' + shop + '/admin/api/2024-01/products.json?limit=250&fields=id,title,variants'
        while url:
            req = urllib.request.Request(url, headers={'X-Shopify-Access-Token': token})
            try:
                with urllib.request.urlopen(req, timeout=20) as r:
                    data = json.loads(r.read())
                    link_header = r.headers.get('Link', '')
            except urllib.error.HTTPError as e:
                body = e.read().decode('utf-8', errors='replace')[:300]
                return jsonify({'ok': False, 'error': 'Shopify HTTP ' + str(e.code) + ' — ' + body + ' | Verifica que el token tenga scope read_products.'})
            except Exception as e:
                return jsonify({'ok': False, 'error': 'Error red Shopify: ' + str(e)})

            for product in data.get('products', []):
                title = str(product.get('title', '') or '').strip()
                for variant in product.get('variants', []):
                    sku_raw = str(variant.get('sku', '') or '').strip().upper()
                    # ⚠ Audit zero-error 2-may-2026: `inventory_quantity` es ON HAND
                    # (incluye committed). Memoria de Sebastián: para MRP/forecast
                    # debe usarse AVAILABLE (= On hand - Committed). Fix completo
                    # requiere segunda API call a /inventory_levels.json con
                    # inventory_item_ids. Pendiente en ROADMAP. Por ahora se usa
                    # On hand y el modelo descuenta pipeline 7d como aproximación.
                    inv_qty_on_hand = int(variant.get('inventory_quantity', 0) or 0)
                    inv_item_id = variant.get('inventory_item_id')  # para fix futuro
                    all_variants.append({
                        'sku': sku_raw, 'titulo': title,
                        'inv_qty': inv_qty_on_hand,
                        'inv_item_id': inv_item_id,
                    })

            next_url = None
            for part in link_header.split(','):
                if 'rel="next"' in part:
                    s = part.find('<') + 1
                    e2 = part.find('>')
                    if s > 0 and e2 > s:
                        next_url = part[s:e2].strip()
            url = next_url

        if not all_variants:
            return jsonify({'ok': False, 'error': 'Shopify no devolvio productos. Tienda sin productos o token sin permiso read_products.'})

        # Sebastián 12-may-2026: fix #D · usar AVAILABLE en lugar de ON HAND.
        # inventory_quantity (paso anterior) = On hand · incluye committed.
        # available (de inventory_levels.json) = lo realmente disponible.
        # Si la segunda API call falla, caemos back a inventory_quantity (mejor algo que nada).
        inv_item_ids = [v['inv_item_id'] for v in all_variants if v.get('inv_item_id')]
        # FIX 1-jun-2026 · SOLO tienda ÁNIMUS LAB (no sumar locations fantasma)
        _loc_id = _shopify_location_id(conn, token, shop)
        avail_map = _fetch_shopify_available(token, shop, inv_item_ids, location_id=_loc_id)
        used_available = bool(avail_map)  # True si Shopify nos dio Available (fix activo)

        conn.execute("UPDATE stock_pt SET estado='Ajustado' WHERE lote_produccion LIKE 'SHOPIFY-%'")
        synced = 0
        skipped = 0
        today = datetime.now().strftime('%Y-%m-%d')

        for v in all_variants:
            sku = v['sku']
            # Preferir Available si está disponible para este inv_item_id;
            # sino fallback a inventory_quantity (On hand).
            iid = v.get('inv_item_id')
            if iid and iid in avail_map:
                qty = max(int(avail_map[iid]), 0)
            else:
                qty = int(v['inv_qty'])
            producto = sku_map.get(sku)
            if not producto:
                prefix = sku.split('-')[0] if '-' in sku else sku[:6]
                producto = sku_map.get(prefix) or v['titulo']
            # FIX 23-may-2026 · auditoría · Bug #4 del 22-may estaba aplicado
            # solo en auto_plan_jobs.job_sync_stock_shopify_diario · este
            # endpoint (botón manual) seguía skippeando qty<=0 · lookup vía
            # sku_producto_map se rompía después · ahora insertar con
            # estado='Agotado' para que velocidad SÍ pueda calcular.
            if qty <= 0:
                skipped += 1
                estado_row = 'Agotado'
                obs = ('Sync Shopify · agotado (Available=0)'
                       if (iid and iid in avail_map)
                       else 'Sync Shopify · agotado (On hand=0)')
            else:
                estado_row = 'Disponible'
                obs = ('Sync Shopify (Available)'
                       if (iid and iid in avail_map)
                       else 'Sync Shopify (On hand)')
            conn.execute(
                "INSERT INTO stock_pt (sku,descripcion,lote_produccion,fecha_produccion,unidades_inicial,unidades_disponible,precio_base,empresa,estado,observaciones) VALUES (?,?,?,?,?,?,0,'ANIMUS',?,?)",
                (sku, producto, 'SHOPIFY-' + today, today, qty, qty, estado_row, obs)
            )
            synced += 1

        _stamp_stock_sync(conn)  # habilita auto-refresh en vivo de Necesidades
        conn.commit()
        return jsonify({
            'ok': True,
            'synced': synced,
            'skipped_zero': skipped,
            'total_variantes': len(all_variants),
            'usado_available': used_available,
            'mensaje': str(synced) + ' SKUs sincronizados desde Shopify' + (' (Available)' if used_available else ' (On hand fallback)'),
        })

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': 'Error interno: ' + str(e), 'trace': traceback.format_exc()[-500:]})
    finally:
        try:
            _liberar_lock_cron(_lock_conn, 'sync_stock_shopify')
        except Exception:
            pass


@bp.route('/api/programacion/debug-stock')
def prog_debug_stock():
    """Debug: stock_pt raw, sku_map y stock calculado por producto.

    SEC-FIX 23-may-2026 · auditoría · era público · exponía inventario
    completo (raw + calculado + mapping).
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()

    # Raw stock_pt
    try:
        raw_pt = conn.execute(
            "SELECT sku, descripcion, unidades_disponible, estado, lote_produccion FROM stock_pt ORDER BY sku"
        ).fetchall()
        raw_pt_list = [{'sku': r[0], 'desc': r[1], 'uds': r[2], 'estado': r[3], 'lote': r[4]} for r in raw_pt]
    except Exception as e:
        raw_pt_list = [{'error': str(e)}]

    # SKU map
    try:
        sku_map_rows = conn.execute("SELECT sku, producto_nombre, activo FROM sku_producto_map").fetchall()
        sku_map_list = [{'sku': r[0], 'producto': r[1], 'activo': r[2]} for r in sku_map_rows]
    except Exception as e:
        sku_map_list = [{'error': str(e)}]

    # Acondicionamiento summary
    try:
        acon = conn.execute(
            "SELECT producto, estado, SUM(unidades_producidas) FROM acondicionamiento GROUP BY producto, estado"
        ).fetchall()
        acon_list = [{'producto': r[0], 'estado': r[1], 'uds': r[2]} for r in acon]
    except Exception as e:
        acon_list = [{'error': str(e)}]

    # Calculated stock
    stock_calc = _get_stock_pt(conn)

    # Formula headers product names
    try:
        fh = conn.execute('SELECT producto_nombre, lote_size_kg FROM formula_headers ORDER BY producto_nombre').fetchall()
        fh_list = [{'producto': r[0], 'lote_kg': r[1]} for r in fh]
    except Exception as e:
        fh_list = [{'error': str(e)}]

    return jsonify({
        'formula_headers': fh_list,
        'stock_pt_raw': raw_pt_list,
        'sku_producto_map': sku_map_list,
        'acondicionamiento': acon_list,
        'stock_calculado': stock_calc,
    })


@bp.route('/api/programacion/velocidad')
def prog_velocidad():
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    # 27-may-2026 · antes int() sin try → ?days=abc tiraba 500 con HTML
    try:
        days = max(1, min(int(request.args.get('days', 60) or 60), 365))
    except (ValueError, TypeError):
        days = 60
    conn = get_db()
    return jsonify(_shopify_velocity(conn, days=days))


@bp.route('/api/programacion/debug-ventas')
def prog_debug_ventas():
    """Diagnostico completo: datos crudos de animus_shopify_orders.

    SEC-FIX 23-may-2026 · auditoría · era público · exponía 5 órdenes
    muestra con sku_items + volumen / fechas (datos comerciales).
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()
    meta = conn.execute(
        "SELECT COUNT(*), MIN(creado_en), MAX(creado_en) FROM animus_shopify_orders"
    ).fetchone()
    since = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')
    rows_60 = conn.execute(
        "SELECT COUNT(*) FROM animus_shopify_orders WHERE creado_en >= ?", (since,)
    ).fetchone()[0]
    samples = conn.execute(
        "SELECT shopify_id, creado_en, sku_items, unidades_total FROM animus_shopify_orders WHERE creado_en >= ? AND sku_items IS NOT NULL ORDER BY creado_en DESC LIMIT 5",
        (since,)
    ).fetchall()
    sample_list = [{'id': r[0], 'fecha': r[1], 'sku_items': r[2], 'uds': r[3]} for r in samples]
    vel = _shopify_velocity(conn, days=60)
    return jsonify({
        'ordenes_total_tabla': meta[0],
        'fecha_min': meta[1],
        'fecha_max': meta[2],
        'ordenes_ultimos_60d': rows_60,
        'muestra_reciente': sample_list,
        'sku_velocity': vel['sku_velocity'],
        'prod_velocity': vel['prod_velocity'],
        'meses_analizados': vel['months_analyzed'],
    })



# ─── Producción programada (local calendar) ────────────────────────────────

@bp.route('/api/programacion/programar', methods=['GET'])
def prog_listar_eventos():
    """List all future production events.

    Enriquecido (30-abr-2026): incluye sala asignada + operarios por fase
    para que el modal de programar producto y otras vistas muestren el
    estado de asignacion sin pegarle a otro endpoint.
    """
    conn = get_db()
    rows = conn.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes, pp.estado,
               pp.observaciones, pp.creado_en,
               pp.area_id, ap.nombre as area_nombre, ap.codigo as area_codigo,
               od.nombre || ' ' || COALESCE(od.apellido,'')  as op_disp,
               oe.nombre || ' ' || COALESCE(oe.apellido,'')  as op_elab,
               oen.nombre || ' ' || COALESCE(oen.apellido,'') as op_env,
               oa.nombre || ' ' || COALESCE(oa.apellido,'')  as op_acon
        FROM produccion_programada pp
        LEFT JOIN areas_planta ap     ON ap.id  = pp.area_id
        LEFT JOIN operarios_planta od  ON od.id  = pp.operario_dispensacion_id
        LEFT JOIN operarios_planta oe  ON oe.id  = pp.operario_elaboracion_id
        LEFT JOIN operarios_planta oen ON oen.id = pp.operario_envasado_id
        LEFT JOIN operarios_planta oa  ON oa.id  = pp.operario_acondicionamiento_id
        ORDER BY pp.fecha_programada
    """).fetchall()
    return jsonify([{
        'id': r[0], 'producto': r[1], 'fecha': r[2],
        'lotes': r[3], 'estado': r[4], 'observaciones': r[5], 'creado_en': r[6],
        'area_id': r[7], 'area_nombre': r[8], 'area_codigo': r[9],
        'operario_dispensacion':      (r[10] or '').strip() or None,
        'operario_elaboracion':       (r[11] or '').strip() or None,
        'operario_envasado':          (r[12] or '').strip() or None,
        'operario_acondicionamiento': (r[13] or '').strip() or None,
    } for r in rows])


@bp.route('/api/programacion/programar', methods=['POST'])
def prog_crear_evento():
    """Create a production event."""
    if 'compras_user' not in session:
        return jsonify({'ok': False, 'error': 'No autorizado'}), 401
    data = request.get_json(force=True, silent=True) or {}
    producto = (data.get('producto') or '').strip()
    fecha    = (data.get('fecha') or '').strip()
    lotes    = int(data.get('lotes') or 1)
    obs      = (data.get('observaciones') or '').strip()

    if not producto or not fecha:
        return jsonify({'ok': False, 'error': 'producto y fecha son requeridos'}), 400

    # Validate date format
    try:
        import datetime as _dt2
        _dt2.date.fromisoformat(fecha)
    except ValueError:
        return jsonify({'ok': False, 'error': 'fecha inválida (use YYYY-MM-DD)'}), 400

    conn = get_db()
    cur = conn.execute(
        """INSERT INTO produccion_programada (producto, fecha_programada, lotes, observaciones)
           VALUES (?, ?, ?, ?)""",
        (producto, fecha, max(1, lotes), obs)
    )
    pid = cur.lastrowid
    # Audit fix 28-may · regla dura: toda mutación de produccion_programada
    # debe auditarse (hueco de trazabilidad tipo 19-may · desaparición sin rastro).
    try:
        audit_log(conn.cursor(), usuario=session.get('compras_user', 'sistema'),
                  accion='CREAR_PRODUCCION_PROGRAMADA', tabla='produccion_programada',
                  registro_id=pid,
                  despues={'producto': producto, 'fecha': fecha, 'lotes': max(1, lotes)})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'id': pid, 'producto': producto, 'fecha': fecha})


@bp.route('/api/programacion/programar/<int:evento_id>', methods=['DELETE'])
def prog_cancelar_evento(evento_id):
    """Cancel a scheduled production event.

    Audit zero-error 2-may-2026: guard contra cancelar producciones que ya
    descontaron inventario. Si fue completada, redirigir a /revertir-completado
    que sí revierte stock. Cancelar una producción completada sin revertir
    dejaría stock fantasma (consumido pero sin razón) → drift de inventario.
    """
    if 'compras_user' not in session:
        return jsonify({'ok': False, 'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT estado, COALESCE(inventario_descontado_at,'') FROM produccion_programada WHERE id=?",
        (evento_id,)
    ).fetchone()
    if not row:
        return jsonify({'ok': False, 'error': 'Producción no encontrada'}), 404
    estado_actual = row[0] or ''
    descontado_at = row[1]
    if estado_actual == 'cancelado':
        return jsonify({'ok': True, 'id': evento_id, 'ya_cancelado': True})
    if estado_actual == 'completado' or descontado_at:
        return jsonify({
            'ok': False,
            'error': 'Producción ya completada · usar /revertir-completado para revertir descuento de inventario',
            'codigo': 'YA_COMPLETADA',
            'inventario_descontado_at': descontado_at or None,
        }), 409
    estado_anterior = estado_actual
    c.execute(
        "UPDATE produccion_programada SET estado='cancelado' WHERE id=?",
        (evento_id,)
    )
    # Fix #4 · 21-may-2026 · liberar SOLs vinculadas a esta producción
    # (canal Pre-Producción · `produccion_checklist`). Antes quedaban
    # huérfanas · Catalina aprobaba compras para producción que ya no existía.
    sols_liberadas = []
    try:
        r_pcl = c.execute(
            """SELECT DISTINCT solicitud_numero
               FROM produccion_checklist
               WHERE produccion_id=? AND COALESCE(solicitud_numero,'') != ''""",
            (evento_id,),
        ).fetchall()
        nums = [r[0] for r in r_pcl if r[0]]
        for num in nums:
            try:
                c.execute(
                    """UPDATE solicitudes_compra
                       SET estado='Cancelada',
                           observaciones = COALESCE(observaciones,'') ||
                                          ' | Producción origen cancelada (id='||?||')'
                       WHERE numero=? AND estado IN ('Pendiente','Aprobada')""",
                    (str(evento_id), num),
                )
                if c.rowcount > 0:
                    sols_liberadas.append(num)
            except Exception:
                continue
    except Exception as e:
        log.warning('liberar SOLs pre-prod fallo: %s', e)
    try:
        audit_log(c, usuario=user, accion='CANCELAR_PRODUCCION',
                  tabla='produccion_programada', registro_id=evento_id,
                  antes={'estado': estado_anterior},
                  despues={'estado': 'cancelado', 'sols_liberadas': sols_liberadas},
                  detalle=f"Canceló producción id={evento_id} (estaba {estado_anterior}) · liberó {len(sols_liberadas)} SOLs")
    except Exception as e:
        log.warning('audit_log CANCELAR_PRODUCCION fallo: %s', e)
    conn.commit()
    return jsonify({'ok': True, 'id': evento_id, 'sols_liberadas': sols_liberadas})


# ─── Planta: catalogo areas + operarios + asignacion (Capa 2) ──────────────
# Sebastian (30-abr-2026): asignar sala fisica + operario por fase a cada
# produccion. Mayerlin fija dispensacion (regla dura). Las salas tienen
# capacidades distintas post-INVIMA (ver migracion 55).

# ── Reporte yield/merma (Mejora A · 12-may-2026) ────────────────────────────
@bp.route('/api/planta/yield-reporte', methods=['GET'])
def planta_yield_reporte():
    """Yield y merma por producto, agrupado por mes (mig 116).

    Query params opcionales:
      desde · YYYY-MM-DD (default: hace 90 días)
      hasta · YYYY-MM-DD (default: hoy)
      producto · filtro exact match (opcional)

    Solo cuenta producciones con kg_real reportado. Las producciones sin
    kg_real aparecen en `pendientes_reportar` para que Sebastián vea cuántas
    faltan llenar. merma_alta_threshold = 5%.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    from datetime import datetime as _dtcls, timezone as _tzcls, timedelta as _td
    # Sebastián 12-may-2026: usar UTC (alineado con SQLite datetime('now', '-5 hours') sin
    # modifier que devuelve UTC). En máquinas con TZ != UTC, date.today() no
    # coincide con date(fin_real_at) y el reporte sale vacío.
    desde = (request.args.get('desde') or '').strip()
    hasta = (request.args.get('hasta') or '').strip()
    producto_filter = (request.args.get('producto') or '').strip()
    hoy_utc = _dtcls.now(_tzcls.utc).date()
    if not desde:
        desde = (hoy_utc - _td(days=90)).isoformat()
    if not hasta:
        hasta = hoy_utc.isoformat()

    conn = get_db()
    where = ["fin_real_at IS NOT NULL",
             "date(fin_real_at) BETWEEN ? AND ?"]
    params = [desde, hasta]
    if producto_filter:
        where.append("producto = ?")
        params.append(producto_filter)

    # Resumen por (producto, mes_yyyy_mm)
    rows = conn.execute(f"""
        SELECT
          producto,
          substr(fin_real_at, 1, 7) AS mes,
          COUNT(*) AS lotes,
          SUM(CASE WHEN kg_real IS NOT NULL THEN 1 ELSE 0 END) AS lotes_con_real,
          SUM(cantidad_kg) AS planeado_kg_total,
          SUM(kg_real) AS real_kg_total,
          AVG(merma_pct) AS merma_pct_avg,
          MIN(merma_pct) AS merma_pct_min,
          MAX(merma_pct) AS merma_pct_max
        FROM produccion_programada
        WHERE {' AND '.join(where)}
        GROUP BY producto, mes
        ORDER BY mes DESC, producto
    """, params).fetchall()

    items = []
    THRESHOLD = 5.0
    for r in rows:
        merma_avg = r['merma_pct_avg']
        items.append({
            'producto': r['producto'],
            'mes': r['mes'],
            'lotes': r['lotes'],
            'lotes_con_kg_real': r['lotes_con_real'] or 0,
            'planeado_kg_total': r['planeado_kg_total'] or 0,
            'real_kg_total': r['real_kg_total'],
            'merma_pct_avg': round(merma_avg, 2) if merma_avg is not None else None,
            'merma_pct_min': round(r['merma_pct_min'], 2) if r['merma_pct_min'] is not None else None,
            'merma_pct_max': round(r['merma_pct_max'], 2) if r['merma_pct_max'] is not None else None,
            'merma_alta': bool(merma_avg is not None and abs(merma_avg) > THRESHOLD),
        })

    # Cuenta de producciones sin kg_real (pendientes de reportar)
    pendientes = conn.execute(f"""
        SELECT COUNT(*) FROM produccion_programada
        WHERE {' AND '.join(where)} AND kg_real IS NULL
    """, params).fetchone()[0]

    # Top 5 outliers (peor merma · solo donde merma > threshold)
    outliers = conn.execute(f"""
        SELECT id, producto, fin_real_at, cantidad_kg, kg_real, merma_pct
        FROM produccion_programada
        WHERE {' AND '.join(where)} AND merma_pct IS NOT NULL
          AND ABS(merma_pct) > ?
        ORDER BY ABS(merma_pct) DESC LIMIT 10
    """, params + [THRESHOLD]).fetchall()

    return jsonify({
        'desde': desde, 'hasta': hasta,
        'producto_filter': producto_filter or None,
        'merma_alta_threshold_pct': THRESHOLD,
        'total_lotes': sum(it['lotes'] for it in items),
        'total_con_kg_real': sum(it['lotes_con_kg_real'] for it in items),
        'pendientes_reportar': pendientes,
        'items': items,
        'outliers': [dict(r) for r in outliers],
    })


@bp.route('/api/planta/areas', methods=['GET'])
def planta_listar_areas():
    """Lista las 5 salas con capacidades y disponibilidad opcional por fecha.

    Query params:
        fecha: YYYY-MM-DD opcional. Si viene, agrega ocupada_por con la
               produccion que tiene asignada esa sala en esa fecha.
        requiere_marmita_ml: opcional. Si viene, marca solo las que tengan
                             marmita >= ese tamano.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    fecha = (request.args.get('fecha') or '').strip()
    req_marmita = request.args.get('requiere_marmita_ml')
    try:
        req_marmita = int(req_marmita) if req_marmita else None
    except (TypeError, ValueError):
        req_marmita = None

    conn = get_db()
    rows = conn.execute("""
        SELECT id, codigo, nombre, puede_producir, puede_envasar,
               marmita_ml, especial, estado, orden
        FROM areas_planta
        WHERE activo=1
        ORDER BY orden, id
    """).fetchall()

    # Si dieron fecha, calcular ocupacion de cada sala ese dia
    ocupacion = {}
    if fecha:
        try:
            date.fromisoformat(fecha)
            for r in conn.execute("""
                SELECT pp.area_id, pp.id, pp.producto, pp.lotes,
                       COALESCE(pp.lotes,1)*COALESCE(fh.lote_size_kg,0) as kg
                FROM produccion_programada pp
                LEFT JOIN formula_headers fh
                    ON UPPER(TRIM(fh.producto_nombre))=UPPER(TRIM(pp.producto))
                WHERE pp.fecha_programada=?
                  AND pp.area_id IS NOT NULL
                  AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado')
            """, (fecha,)):
                ocupacion.setdefault(r[0], []).append({
                    'produccion_id': r[1], 'producto': r[2],
                    'lotes': r[3], 'kg': r[4],
                })
        except ValueError:
            pass

    out = []
    for r in rows:
        area = {
            'id': r[0], 'codigo': r[1], 'nombre': r[2],
            'puede_producir': bool(r[3]), 'puede_envasar': bool(r[4]),
            'marmita_ml': r[5], 'especial': r[6],
            'estado': r[7], 'orden': r[8],
        }
        ocup = ocupacion.get(r[0], [])
        area['ocupada_por'] = ocup
        area['libre_en_fecha'] = (len(ocup) == 0) if fecha else None
        if req_marmita is not None:
            area['cumple_marmita'] = (r[5] is not None and r[5] >= req_marmita)
        out.append(area)
    return jsonify({'areas': out, 'fecha': fecha or None})


@bp.route('/api/planta/areas/<int:area_id>/estado', methods=['PATCH'])
def planta_actualizar_estado_area(area_id):
    """Cambia el estado de una sala: libre / ocupada / sucia / limpiando.
    Lo usa quien marca que la senora del aseo ya termino, o que una sala
    quedo sucia despues de produccion."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    # BUG-8 fix · 20-may-2026 OLA 1: cualquier user logueado podía marcar
    # libre/sucia/ocupada una sala · contadora/comercial/marketing podían
    # marcar PROD1 'libre' con producción activa → contaminación. Solo
    # PLANTA_USERS o ADMIN_USERS pueden tocar estado de salas.
    try:
        from config import PLANTA_USERS, ADMIN_USERS
        _permitidos = set(ADMIN_USERS) | set(PLANTA_USERS)
        if user not in _permitidos:
            return jsonify({'error': 'Solo planta/admin pueden cambiar estado de sala'}), 403
    except ImportError:
        pass
    data = request.get_json(force=True, silent=True) or {}
    nuevo = (data.get('estado') or '').strip().lower()
    if nuevo not in ('libre', 'ocupada', 'sucia', 'limpiando'):
        return jsonify({'error': 'estado invalido'}), 400
    conn = get_db()
    # Capturar estado anterior para audit
    prev_row = conn.execute(
        "SELECT estado FROM areas_planta WHERE id=? AND activo=1",
        (area_id,)
    ).fetchone()
    if not prev_row:
        return jsonify({'error': 'sala no encontrada'}), 404
    estado_anterior = prev_row[0]
    # INVIMA-FIX · 21-may-2026 · bloquear bypass de despeje
    # Transición sucia → libre solo via /api/planta/despeje-linea (checklist 5 ítems)
    # · admin puede forzar con flag forzar_sin_despeje=true (queda en audit)
    if estado_anterior == 'sucia' and nuevo == 'libre':
        if not data.get('forzar_sin_despeje'):
            return jsonify({
                'error': 'Transición sucia→libre requiere despeje firmado · usar /api/planta/despeje-linea',
                'codigo': 'DESPEJE_REQUERIDO',
                'hint': 'Si excepción autorizada admin, enviar forzar_sin_despeje=true',
            }), 409
        if user not in ADMIN_USERS:
            return jsonify({
                'error': 'Forzar despeje solo admin · contactá Sebastián',
                'codigo': 'FORZAR_SOLO_ADMIN',
            }), 403
    cur = conn.execute(
        "UPDATE areas_planta SET estado=? WHERE id=? AND activo=1",
        (nuevo, area_id)
    )
    conn.commit()
    if cur.rowcount == 0:
        return jsonify({'error': 'sala no encontrada'}), 404
    try:
        audit_log(conn.cursor(), usuario=user,
                  accion='ACTUALIZAR_ESTADO_AREA',
                  tabla='areas_planta', registro_id=area_id,
                  antes={'estado': estado_anterior},
                  despues={'estado': nuevo},
                  detalle=f"Cambió estado de sala {area_id}: {estado_anterior} → {nuevo}")
    except Exception as _e:
        logging.getLogger('programacion').warning(
            f'audit ACTUALIZAR_ESTADO_AREA fallo: {_e}')
    return jsonify({'ok': True, 'id': area_id, 'estado': nuevo,
                    'estado_anterior': estado_anterior})


# ════════════════════════════════════════════════════════════════════════
# RÓTULO VIRTUAL DE LIMPIEZA · PRD-PRO-002-F02 (Estado de Limpieza de
# Áreas/Equipos) · Sebastián 6-jun-2026.
#
# Reemplaza el rótulo físico que se imprimía y diligenciaba a mano. Fluye
# SOLO con la producción: el estado físico (Limpio/En uso/Sucio) ya vive en
# areas_planta.estado (libre/ocupada/sucia) — NO se duplica acá (M5/M9). Esta
# capa agrega el REGISTRO F02 por ciclo de limpieza (rotulos_limpieza), un
# snapshot inmutable Part 11 con dos firmas: operario realiza · Calidad
# verifica. La liberación física sucia→libre usa la ruta ÚNICA
# liberar_sala_con_despeje (M3 · auto_plan.py).
# ════════════════════════════════════════════════════════════════════════

# Estado físico de la sala → etiqueta del rótulo F02.
_ESTADO_ROTULO = {
    'libre': 'Limpio', 'ocupada': 'En uso',
    'sucia': 'Sucio', 'limpiando': 'En limpieza',
}
# Las salas legacy de programación (PROD1..PROD4) no comparten código con el
# catálogo de equipos (FAB1/FAB2/FAB3/ENV2). Alias para listar sus equipos.
_SALA_EQUIPO_ALIAS = {
    'PROD1': 'FAB1', 'PROD2': 'FAB2', 'PROD3': 'FAB3',
}


def _now_co():
    """Timestamp local Colombia (UTC-5) calculado en Python · PG-safe (no
    datetime() en DML)."""
    return (datetime.now(timezone.utc) - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')


def _equipos_de_area(c, area_codigo):
    """Equipos activos de un área (catálogo equipos_planta). Resuelve el alias
    sala-legacy→área-equipos para PROD1..PROD4."""
    alias = _SALA_EQUIPO_ALIAS.get(area_codigo, area_codigo)
    try:
        rows = c.execute(
            """SELECT codigo, nombre, tipo FROM equipos_planta
               WHERE area_codigo IN (?, ?) AND activo=1
               ORDER BY codigo""",
            (area_codigo, alias),
        ).fetchall()
    except Exception:
        return []
    return [{'codigo': r[0], 'nombre': r[1], 'tipo': r[2] or ''} for r in rows]


def _lote_de_produccion(c, produccion_id):
    """Lote real desde el EBR (ebr_ejecuciones.lote) si existe. Best-effort."""
    if not produccion_id:
        return ''
    try:
        r = c.execute(
            "SELECT lote FROM ebr_ejecuciones WHERE produccion_id=? ORDER BY id DESC LIMIT 1",
            (produccion_id,),
        ).fetchone()
        return (r[0] if r and r[0] else '') or ''
    except Exception:
        return ''


def _rotulo_derivar(c, area_id):
    """Deriva en vivo los datos del rótulo de un área (defaults antes de
    firmar). Devuelve dict o None si el área no existe."""
    a = c.execute(
        "SELECT id, codigo, nombre, estado FROM areas_planta WHERE id=? AND activo=1",
        (area_id,),
    ).fetchone()
    if not a:
        return None
    estado_fisico = a[3] or 'libre'
    # Producto a elaborar = producción en curso o próxima en el área.
    prod_act = c.execute(
        """SELECT id, producto FROM produccion_programada
           WHERE area_id=? AND COALESCE(estado,'programado') NOT IN ('completado','cancelado')
           ORDER BY (inicio_real_at IS NULL), fecha_programada ASC LIMIT 1""",
        (area_id,),
    ).fetchone()
    # Producto anterior = último lote terminado físicamente en el área.
    prod_prev = c.execute(
        """SELECT id, producto FROM produccion_programada
           WHERE area_id=? AND fin_real_at IS NOT NULL
           ORDER BY fin_real_at DESC LIMIT 1""",
        (area_id,),
    ).fetchone()
    return {
        'area_id': a[0], 'area_codigo': a[1], 'area_nombre': a[2],
        'estado_fisico': estado_fisico,
        'estado': _ESTADO_ROTULO.get(estado_fisico, estado_fisico),
        'produccion_id': (prod_act[0] if prod_act else None),
        'producto_elaborar': (prod_act[1] if prod_act else ''),
        'lote_elaborar': _lote_de_produccion(c, prod_act[0]) if prod_act else '',
        'producto_anterior': (prod_prev[1] if prod_prev else ''),
        'lote_anterior': _lote_de_produccion(c, prod_prev[0]) if prod_prev else '',
        'equipos': _equipos_de_area(c, a[1]),
    }


def _rotulo_abierto(c, area_id):
    """Fila del ciclo de limpieza ABIERTO (aún sin verificar) del área, o None."""
    try:
        r = c.execute(
            """SELECT id, sanitizante, detergente, equipos_json, estado,
                      realizado_por, realizado_at, verificado_por, verificado_at,
                      producto_elaborar, lote_elaborar, producto_anterior,
                      lote_anterior, observaciones
               FROM rotulos_limpieza
               WHERE area_id=? AND COALESCE(verificado_at,'')=''
               ORDER BY id DESC LIMIT 1""",
            (area_id,),
        ).fetchone()
    except Exception:
        return None
    if not r:
        return None
    return {
        'id': r[0], 'sanitizante': r[1], 'detergente': r[2],
        'equipos_json': r[3] or '', 'estado': r[4],
        'realizado_por': r[5] or '', 'realizado_at': r[6] or '',
        'verificado_por': r[7] or '', 'verificado_at': r[8] or '',
        'producto_elaborar': r[9] or '', 'lote_elaborar': r[10] or '',
        'producto_anterior': r[11] or '', 'lote_anterior': r[12] or '',
        'observaciones': r[13] or '',
    }


@bp.route('/api/planta/rotulo-limpieza/<int:area_id>', methods=['GET'])
def planta_rotulo_limpieza_get(area_id):
    """Vista viva del rótulo de limpieza F02 de un área: estado físico actual
    (Limpio/En uso/Sucio), producto a elaborar+lote, producto anterior+lote,
    equipos del área, y el ciclo de limpieza abierto si existe."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = (session.get('compras_user') or '')
    conn = get_db(); c = conn.cursor()
    base = _rotulo_derivar(c, area_id)
    if not base:
        return jsonify({'error': 'área no encontrada'}), 404
    abierto = _rotulo_abierto(c, area_id)
    # Permisos para los botones de acción.
    try:
        es_planta = user in (set(ADMIN_USERS) | set(PLANTA_USERS))
        es_calidad = user in (set(ADMIN_USERS) | set(CALIDAD_USERS))
    except Exception:
        es_planta = es_calidad = False
    # Puede REALIZAR limpieza: la sala está sucia (o en limpieza) y es planta.
    puede_realizar = es_planta and base['estado_fisico'] in ('sucia', 'limpiando')
    # Puede VERIFICAR: hay limpieza realizada sin verificar y es calidad.
    puede_verificar = bool(
        es_calidad and abierto and abierto.get('realizado_at')
        and not abierto.get('verificado_at'))
    return jsonify({
        'ok': True,
        'rotulo': base,
        'ciclo': abierto,
        'puede_realizar': puede_realizar,
        'puede_verificar': puede_verificar,
        'es_planta': es_planta,
        'es_calidad': es_calidad,
        'sanitizantes_sugeridos': ['Alcohol 70%', 'Amonio Cuaternario', 'Hipoclorito 200ppm'],
        'detergentes_sugeridos': ['Detergente Neutro Industrial', 'Desengrasante alcalino'],
    })


@bp.route('/api/planta/rotulos-limpieza', methods=['GET'])
def planta_rotulos_limpieza_lista():
    """Lista las áreas limpiables (las 7 oficiales: las de producción + Dispensación
    + Acondicionamiento) con su estado actual · para la sub-pestaña de rótulos."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute(
        """SELECT id, codigo, nombre, COALESCE(estado,'libre'), orden
           FROM areas_planta
           WHERE activo=1 AND (tipo='produccion' OR codigo IN ('DISP','ACOND'))
           ORDER BY orden, codigo"""
    ).fetchall()
    areas = []
    for r in rows:
        aid, codigo, nombre, estado = r[0], r[1], r[2], r[3]
        prod = c.execute(
            """SELECT producto FROM produccion_programada
               WHERE area_id=? AND COALESCE(estado,'programado') NOT IN ('completado','cancelado')
               ORDER BY (inicio_real_at IS NULL), fecha_programada ASC LIMIT 1""",
            (aid,)).fetchone()
        areas.append({
            'id': aid, 'codigo': codigo, 'nombre': nombre, 'estado': estado,
            'estado_rotulo': _ESTADO_ROTULO.get(estado, estado),
            'producto': (prod[0] if prod else ''),
        })
    return jsonify({'ok': True, 'areas': areas})


@bp.route('/api/planta/rotulo-limpieza/<int:area_id>/realizar', methods=['POST'])
def planta_rotulo_limpieza_realizar(area_id):
    """Operario registra que EJECUTÓ la limpieza del área (sucia → limpiando).
    Fotografía (snapshot Part 11) producto/lote, sanitizante, detergente y los
    equipos limpiados. NO libera la sala — eso lo hace Calidad al verificar.

    Body: {sanitizante?, detergente?, equipos?:[codigos], observaciones?,
           producto_elaborar?, lote_elaborar?, producto_anterior?, lote_anterior?}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = (session.get('compras_user') or '')
    try:
        if user not in (set(ADMIN_USERS) | set(PLANTA_USERS)):
            return jsonify({'error': 'Solo planta/admin pueden registrar limpieza'}), 403
    except Exception:
        pass
    conn = get_db(); c = conn.cursor()
    base = _rotulo_derivar(c, area_id)
    if not base:
        return jsonify({'error': 'área no encontrada'}), 404
    if base['estado_fisico'] not in ('sucia', 'limpiando'):
        return jsonify({
            'error': f"El área está '{base['estado']}' · solo se registra limpieza "
                     f"cuando está Sucio (tras producción)",
            'codigo': 'AREA_NO_SUCIA',
            'estado': base['estado'],
        }), 409
    body = request.get_json(silent=True) or {}
    sanitizante = (body.get('sanitizante') or 'Alcohol 70%').strip()[:120]
    detergente = (body.get('detergente') or 'Detergente Neutro Industrial').strip()[:120]
    obs = (body.get('observaciones') or '').strip()[:500]
    equipos_sel = body.get('equipos') or []
    if not isinstance(equipos_sel, list):
        equipos_sel = []
    equipos_json = json.dumps([str(x)[:40] for x in equipos_sel][:60], ensure_ascii=False)
    # Snapshot: lo que el operario confirma/edita o, si no, lo derivado.
    prod_elab = (body.get('producto_elaborar') or base['producto_elaborar'] or '').strip()[:200]
    lote_elab = (body.get('lote_elaborar') or base['lote_elaborar'] or '').strip()[:80]
    prod_prev = (body.get('producto_anterior') or base['producto_anterior'] or '').strip()[:200]
    lote_prev = (body.get('lote_anterior') or base['lote_anterior'] or '').strip()[:80]
    ahora = _now_co()
    abierto = _rotulo_abierto(c, area_id)
    if abierto:
        rid = abierto['id']
        c.execute(
            """UPDATE rotulos_limpieza
                 SET sanitizante=?, detergente=?, equipos_json=?, estado='realizado',
                     realizado_por=?, realizado_at=?, producto_elaborar=?,
                     lote_elaborar=?, producto_anterior=?, lote_anterior=?,
                     observaciones=?, actualizado_en=?
               WHERE id=?""",
            (sanitizante, detergente, equipos_json, user, ahora, prod_elab,
             lote_elab, prod_prev, lote_prev, obs, ahora, rid),
        )
    else:
        c.execute(
            """INSERT INTO rotulos_limpieza
                 (area_id, area_codigo, produccion_id, producto_elaborar,
                  lote_elaborar, producto_anterior, lote_anterior, sanitizante,
                  detergente, equipos_json, estado, realizado_por, realizado_at,
                  observaciones, creado_en, actualizado_en)
               VALUES (?,?,?,?,?,?,?,?,?,?,'realizado',?,?,?,?,?)""",
            (area_id, base['area_codigo'], base['produccion_id'], prod_elab,
             lote_elab, prod_prev, lote_prev, sanitizante, detergente,
             equipos_json, user, ahora, obs, ahora, ahora),
        )
        rid = c.lastrowid
    # Estado físico: sucia → limpiando (no libera). Idempotente.
    estado_prev = base['estado_fisico']
    c.execute(
        "UPDATE areas_planta SET estado='limpiando' WHERE id=? AND estado IN ('sucia','limpiando')",
        (area_id,),
    )
    try:
        c.execute(
            """INSERT INTO area_eventos
                 (area_id, tipo, estado_anterior, estado_nuevo, usuario, nota)
               VALUES (?,?,?,?,?,?)""",
            (area_id, 'limpieza_realizada', estado_prev, 'limpiando', user,
             f'Limpieza ejecutada · {sanitizante} · rótulo #{rid}'),
        )
    except Exception:
        pass
    try:
        audit_log(c, usuario=user, accion='ROTULO_LIMPIEZA_REALIZAR',
                  tabla='rotulos_limpieza', registro_id=rid,
                  antes={'estado_area': estado_prev},
                  despues={'estado_area': 'limpiando', 'sanitizante': sanitizante,
                           'equipos': equipos_sel[:20], 'producto_anterior': prod_prev,
                           'producto_elaborar': prod_elab},
                  detalle=f"Limpieza registrada en {base['area_codigo']} por {user}")
    except Exception as _e:
        logging.getLogger('programacion').warning(f'audit ROTULO_LIMPIEZA_REALIZAR: {_e}')
    conn.commit()
    return jsonify({'ok': True, 'rotulo_id': rid, 'estado': 'En limpieza',
                    'mensaje': 'Limpieza registrada · pendiente verificación de Calidad'})


@bp.route('/api/planta/rotulo-limpieza/<int:area_id>/verificar', methods=['POST'])
def planta_rotulo_limpieza_verificar(area_id):
    """Calidad VERIFICA la limpieza y libera el área (limpiando → libre).
    Requiere e-firma Part 11 (meaning='verifica' sobre el rótulo). Usa la ruta
    ÚNICA de liberación con despeje (M3) y cierra el registro F02 (inmutable).

    Body: {signature_id (req), observaciones?}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = (session.get('compras_user') or '')
    try:
        if user not in (set(ADMIN_USERS) | set(CALIDAD_USERS)):
            return jsonify({'error': 'Solo Calidad/admin pueden verificar la limpieza'}), 403
    except Exception:
        pass
    conn = get_db(); c = conn.cursor()
    base = _rotulo_derivar(c, area_id)
    if not base:
        return jsonify({'error': 'área no encontrada'}), 404
    abierto = _rotulo_abierto(c, area_id)
    if not abierto or not abierto.get('realizado_at'):
        return jsonify({
            'error': 'No hay limpieza realizada pendiente de verificar en esta área',
            'codigo': 'SIN_LIMPIEZA_REALIZADA',
        }), 409
    rid = abierto['id']
    body = request.get_json(silent=True) or {}
    signature_id = body.get('signature_id')
    if not signature_id:
        return jsonify({
            'error': "Firma requerida · primero POST /api/sign con "
                     "{record_table:'rotulos_limpieza', record_id:'%s', meaning:'revisa'}" % rid,
            'codigo': 'FIRMA_REQUERIDA',
            'rotulo_id': rid,
        }), 400
    try:
        from blueprints.inventario import _validar_e_sign
    except ImportError:
        from inventario import _validar_e_sign
    try:
        firma_ok = _validar_e_sign(
            c, signature_id, record_table='rotulos_limpieza',
            record_id=rid, meaning='revisa', signer_username=user)
    except Exception:
        firma_ok = False
    if not firma_ok:
        return jsonify({
            'error': 'La firma no corresponde a una verificación de este rótulo por vos',
            'codigo': 'FIRMA_INVALIDA',
        }), 400
    obs = (body.get('observaciones') or '').strip()[:500]
    # Liberación física por la ruta ÚNICA (M3): despeje firmado + sala libre.
    try:
        from blueprints.auto_plan import liberar_sala_con_despeje
    except ImportError:
        from auto_plan import liberar_sala_con_despeje
    try:
        checklist_id, area_codigo, estado_prev = liberar_sala_con_despeje(
            c, area_id, user,
            obs=(obs or f'Verificación de limpieza · rótulo #{rid}'),
            rotulo_id=rid, verificado_por=user)
    except ValueError:
        return jsonify({'error': 'área no encontrada'}), 404
    ahora = _now_co()
    c.execute(
        """UPDATE rotulos_limpieza
             SET estado='verificado', verificado_por=?, verificado_at=?,
                 verificado_sign_id=?, despeje_checklist_id=?, actualizado_en=?
           WHERE id=?""",
        (user, ahora, int(signature_id), checklist_id, ahora, rid),
    )
    try:
        c.execute(
            """INSERT INTO area_eventos
                 (area_id, tipo, estado_anterior, estado_nuevo, usuario, nota)
               VALUES (?,?,?,?,?,?)""",
            (area_id, 'limpieza_verificada', estado_prev, 'libre', user,
             f'Limpieza verificada por Calidad · rótulo #{rid} · despeje #{checklist_id}'),
        )
    except Exception:
        pass
    try:
        audit_log(c, usuario=user, accion='ROTULO_LIMPIEZA_VERIFICAR',
                  tabla='rotulos_limpieza', registro_id=rid,
                  antes={'estado_area': estado_prev},
                  despues={'estado_area': 'libre', 'signature_id': int(signature_id),
                           'despeje_checklist_id': checklist_id},
                  detalle=f"Limpieza verificada y área {area_codigo} liberada por {user}")
    except Exception as _e:
        logging.getLogger('programacion').warning(f'audit ROTULO_LIMPIEZA_VERIFICAR: {_e}')
    conn.commit()
    return jsonify({'ok': True, 'rotulo_id': rid, 'estado': 'Limpio',
                    'checklist_id': checklist_id,
                    'mensaje': f'Limpieza verificada · {area_codigo} liberada (Limpio)'})


def _persona_corta(c, username):
    """username → 'Nombre Completo · Cargo' desde usuarios_identidad."""
    if not username:
        return ''
    try:
        r = c.execute(
            "SELECT nombre_completo, cargo FROM usuarios_identidad WHERE username=?",
            (username,),
        ).fetchone()
    except Exception:
        r = None
    if not r:
        return username
    nom = (r[0] or username).strip()
    cargo = (r[1] or '').strip()
    return f'{nom} · {cargo}' if cargo else nom


@bp.route('/planta/rotulo-limpieza/<int:area_id>/pdf', methods=['GET'])
def planta_rotulo_limpieza_pdf(area_id):
    """Rótulo imprimible PRD-PRO-002-F02 (Estado de Limpieza de Áreas/Equipos).
    Renderiza el ciclo de limpieza vigente del área (snapshot inmutable) sobre
    el estado físico actual. HTML puro imprimible (sin scripts)."""
    if 'compras_user' not in session:
        from flask import redirect
        return redirect('/login?next=/planta/rotulo-limpieza/%d/pdf' % area_id)
    from flask import Response
    from html import escape as _e
    conn = get_db(); c = conn.cursor()
    base = _rotulo_derivar(c, area_id)
    if not base:
        return Response('<h1>Área no encontrada</h1>', mimetype='text/html', status=404)
    # Último ciclo (abierto o cerrado) para mostrar el registro F02 vigente.
    rot = c.execute(
        """SELECT producto_elaborar, lote_elaborar, producto_anterior, lote_anterior,
                  sanitizante, detergente, equipos_json, realizado_por, realizado_at,
                  verificado_por, verificado_at
           FROM rotulos_limpieza WHERE area_id=? ORDER BY id DESC LIMIT 1""",
        (area_id,),
    ).fetchone()
    if rot:
        (prod_elab, lote_elab, prod_prev, lote_prev, sanit, deterg, eq_json,
         realizado_por, realizado_at, verificado_por, verificado_at) = rot
    else:
        prod_elab, lote_elab = base['producto_elaborar'], base['lote_elaborar']
        prod_prev, lote_prev = base['producto_anterior'], base['lote_anterior']
        sanit, deterg, eq_json = 'Alcohol 70%', 'Detergente Neutro Industrial', ''
        realizado_por = realizado_at = verificado_por = verificado_at = ''
    try:
        equipos_lst = json.loads(eq_json) if eq_json else []
    except Exception:
        equipos_lst = []
    # Etiqueta de área/equipo: si el ciclo limpió UN solo equipo, mostrarlo.
    area_label = f"{base['area_nombre']} · {base['area_codigo']}"
    equipos_txt = ', '.join(str(x) for x in equipos_lst) if equipos_lst else '—'
    estado_fisico = base['estado_fisico']
    realizado_full = _persona_corta(c, realizado_por)
    verificado_full = _persona_corta(c, verificado_por)

    def _chip(label, activo):
        cls = 'chip on' if activo else 'chip off'
        return f'<span class="{cls}">{_e(label)}</span>'

    estado_chips = (_chip('Limpio', estado_fisico == 'libre')
                    + _chip('En uso', estado_fisico == 'ocupada')
                    + _chip('Sucio', estado_fisico in ('sucia', 'limpiando')))

    def _row(lbl, val, num=False):
        vcls = ' class="num"' if num else ''
        return f'<tr><td class="k">{_e(lbl)}</td><td{vcls}>{_e(str(val) or "—")}</td></tr>'

    html = f'''<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Rótulo de Limpieza F02 · {_e(base['area_codigo'])}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  :root{{--ink:#18181b;--soft:#3f3f46;--mute:#71717a;--line:#e4e4e7;--violet:#6d28d9;--violet-d:#4c1d95;--pale:#f5f3ff}}
  *{{box-sizing:border-box}}
  body{{font-family:'Inter',system-ui,Arial,sans-serif;margin:0;padding:28px;background:#f4f4f7;color:var(--ink);-webkit-font-smoothing:antialiased}}
  .sheet{{max-width:760px;margin:0 auto;background:#fff;border:1px solid var(--line);border-radius:14px;overflow:hidden;box-shadow:0 1px 2px rgba(24,24,27,.05),0 12px 28px rgba(24,24,27,.08)}}
  .accent{{height:5px;background:linear-gradient(90deg,#a78bfa,var(--violet))}}
  .top{{display:flex;justify-content:space-between;align-items:flex-start;gap:18px;padding:20px 26px 14px}}
  .brand{{display:flex;align-items:center;gap:12px}}
  .mark{{width:46px;height:46px;border-radius:13px;flex-shrink:0;display:inline-flex;align-items:center;justify-content:center;background:linear-gradient(135deg,#a78bfa,var(--violet));box-shadow:0 4px 14px rgba(109,40,217,.22)}}
  .brand .co{{font-size:16px;font-weight:800;letter-spacing:-.3px;line-height:1.1}}
  .brand .sub{{font-size:11.5px;color:var(--mute);margin-top:2px;font-weight:500}}
  .ctrl{{font-size:11px;color:var(--soft);text-align:right;line-height:1.7;background:var(--pale);border:1px solid #ede9fe;border-radius:10px;padding:9px 14px}}
  .ctrl b{{color:var(--violet-d);font-weight:700}}
  .title{{text-align:center;padding:6px 26px 16px}}
  .title h1{{margin:0;font-size:19px;font-weight:800;letter-spacing:-.3px;color:var(--ink);text-transform:uppercase}}
  .title .k{{font-size:11.5px;color:var(--mute);margin-top:4px;font-weight:600;letter-spacing:.4px}}
  .estado{{text-align:center;padding:6px 20px 20px;border-bottom:1px solid var(--line)}}
  .estado .elbl{{font-size:11px;font-weight:700;color:var(--mute);letter-spacing:.6px;margin-bottom:10px}}
  .chip{{display:inline-block;padding:9px 20px;margin:0 5px;border-radius:10px;font-weight:700;font-size:14px;letter-spacing:.3px}}
  .chip.on{{background:var(--violet);color:#fff;box-shadow:0 4px 14px rgba(109,40,217,.18)}}
  .chip.off{{border:1px solid var(--line);color:var(--mute)}}
  table{{width:100%;border-collapse:collapse}}
  td{{padding:11px 16px;border-bottom:1px solid var(--line);vertical-align:middle;font-size:14px}}
  td.k{{width:34%;color:var(--mute);font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.4px;background:#fafafa}}
  .num{{font-variant-numeric:tabular-nums;font-weight:600}}
  .firmas{{display:flex;border-top:1px solid var(--line)}}
  .firma{{flex:1;padding:18px 22px 22px}}
  .firma+.firma{{border-left:1px solid var(--line)}}
  .firma .l{{font-size:11px;font-weight:700;color:var(--mute);text-transform:uppercase;letter-spacing:.4px}}
  .firma .v{{font-size:15px;font-weight:600;margin-top:18px;border-top:1px solid var(--ink);padding-top:6px}}
  .firma .f{{font-size:11.5px;color:var(--mute);margin-top:3px;font-variant-numeric:tabular-nums}}
  .printbar{{text-align:center;margin-top:18px}}
  .printbtn{{display:inline-flex;align-items:center;gap:8px;padding:11px 26px;background:var(--violet);color:#fff;text-decoration:none;border:none;border-radius:10px;font-weight:600;font-size:14px;font-family:'Inter';cursor:pointer;box-shadow:0 4px 14px rgba(109,40,217,.22)}}
  @media print{{ body{{padding:0;background:#fff}} .sheet{{box-shadow:none;border:none}} .printbar{{display:none}} }}
</style></head><body>
<div class="sheet">
  <div class="accent"></div>
  <div class="top">
    <div class="brand">
      <span class="mark"><svg viewBox="0 0 32 32" width="28" height="28" fill="none" stroke="#fff" xmlns="http://www.w3.org/2000/svg"><circle cx="16" cy="12" r="3" fill="#fff"/><path d="M 5 19 Q 16 17, 27 19" stroke-width="1.6" stroke-linecap="round" opacity=".7"/><path d="M 5 23 Q 16 21, 27 23" stroke-width="1.6" stroke-linecap="round" opacity=".4"/></svg></span>
      <div><div class="co">ESPAGIRIA Laboratorio SAS</div><div class="sub">ÁNIMUS Lab · Producción</div></div>
    </div>
    <div class="ctrl"><b>Código:</b> PRD-PRO-002-F02<br><b>Versión:</b> 02 &nbsp;·&nbsp; <b>Página:</b> 1 de 1<br><b>Vigencia:</b> 9-Abr-2026 a 8-Abr-2029</div>
  </div>
  <div class="title">
    <h1>Estado de Limpieza de Áreas / Equipos</h1>
    <div class="k">Registro de verificación previo a fabricación · BPM / INVIMA · 21 CFR Part 11</div>
  </div>
  <div class="estado">
    <div class="elbl">ESTADO</div>
    {estado_chips}
  </div>
  <table>
    {_row('Área o equipo · código', area_label)}
    {_row('Equipos', equipos_txt)}
    {_row('Producto a elaborar', prod_elab)}
    {_row('Lote', lote_elab, num=True)}
    {_row('Sanitizante', sanit)}
    {_row('Detergente', deterg)}
    {_row('Producto anterior', prod_prev)}
    {_row('Lote anterior', lote_prev, num=True)}
  </table>
  <div class="firmas">
    <div class="firma">
      <div class="l">Realizado por (Operario)</div>
      <div class="v">{_e(realizado_full or '—')}</div>
      <div class="f">Fecha: {_e(realizado_at or '—')}{' · firma electrónica ✔' if realizado_full else ''}</div>
    </div>
    <div class="firma">
      <div class="l">Verificado por (Calidad)</div>
      <div class="v">{_e(verificado_full or '—')}</div>
      <div class="f">Fecha: {_e(verificado_at or '—')}{' · firma electrónica ✔' if verificado_full else ''}</div>
    </div>
  </div>
</div>
<div class="printbar">
  <button class="printbtn" onclick="window.print()">🖨 Imprimir / Guardar PDF</button>
</div>
</body></html>'''
    return Response(html, mimetype='text/html')


@bp.route('/api/planta/operarios', methods=['GET', 'POST'])
def planta_listar_operarios():
    """GET: lista crew activo (Mayerlin fija dispensación, Luis Enrique jefe).
    POST: crea un nuevo operario. Solo admin (sebastian/alejandro).

    Body POST:
      nombre (req), apellido, rol_predeterminado (todero/envasado/
      acondicionamiento/dispensacion/jefe), fija_en_dispensacion (bool),
      es_jefe_produccion (bool).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = (session.get('compras_user') or '').lower()
    conn = get_db(); c = conn.cursor()

    if request.method == 'POST':
        try:
            from blueprints.compras import ADMIN_USERS as _ADMIN
        except Exception:
            _ADMIN = ('sebastian', 'alejandro')
        if user not in {u.lower() for u in _ADMIN}:
            return jsonify({'error': 'Solo admin crea operarios'}), 403
        d = request.get_json(force=True, silent=True) or {}
        nombre = (d.get('nombre') or '').strip()
        if not nombre:
            return jsonify({'error': 'nombre requerido'}), 400
        rol = (d.get('rol_predeterminado') or 'todero').strip().lower()
        if rol not in ('dispensacion', 'envasado', 'acondicionamiento', 'todero', 'jefe'):
            rol = 'todero'
        fija = 1 if d.get('fija_en_dispensacion') else 0
        jefe = 1 if d.get('es_jefe_produccion') else 0
        c.execute("""INSERT INTO operarios_planta
            (nombre, apellido, rol_predeterminado, fija_en_dispensacion, es_jefe_produccion, activo)
            VALUES (?,?,?,?,?,1)""",
            (nombre, (d.get('apellido') or '').strip() or None, rol, fija, jefe))
        conn.commit()
        return jsonify({'ok': True, 'id': c.lastrowid}), 201

    # GET: incluir tambien inactivos si admin lo pide via ?incluir_inactivos=1
    incluir = request.args.get('incluir_inactivos') == '1'
    sql = """SELECT id, nombre, apellido, rol_predeterminado,
                    fija_en_dispensacion, es_jefe_produccion, activo
             FROM operarios_planta"""
    if not incluir:
        sql += " WHERE activo=1"
    sql += " ORDER BY activo DESC, es_jefe_produccion DESC, fija_en_dispensacion DESC, nombre"
    rows = c.execute(sql).fetchall()
    out = [{
        'id': r[0],
        'nombre': r[1],
        'apellido': r[2] or '',
        'nombre_completo': (r[1] + ' ' + (r[2] or '')).strip(),
        'rol': r[3] or '',
        'fija_dispensacion': bool(r[4]),
        'es_jefe': bool(r[5]),
        'activo': bool(r[6]),
    } for r in rows]
    return jsonify({'operarios': out, 'total': len(out)})


@bp.route('/api/planta/operarios/<int:op_id>', methods=['PATCH', 'DELETE'])
def planta_actualizar_operario(op_id):
    """PATCH: edita campos del operario (nombre, apellido, rol, flags).
    DELETE: soft delete — marca activo=0 (no borra historial de producciones).
    Solo admin."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = (session.get('compras_user') or '').lower()
    try:
        from blueprints.compras import ADMIN_USERS as _ADMIN
    except Exception:
        _ADMIN = ('sebastian', 'alejandro')
    if user not in {u.lower() for u in _ADMIN}:
        return jsonify({'error': 'Solo admin edita operarios'}), 403
    conn = get_db(); c = conn.cursor()

    if request.method == 'DELETE':
        cur = c.execute("UPDATE operarios_planta SET activo=0 WHERE id=?", (op_id,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify({'ok': True, 'id': op_id, 'desactivado': True})

    # PATCH
    d = request.get_json(force=True, silent=True) or {}
    sets = []; params = []
    if 'nombre' in d:
        n = (d['nombre'] or '').strip()
        if not n: return jsonify({'error': 'nombre vacio'}), 400
        sets.append('nombre=?'); params.append(n)
    if 'apellido' in d:
        sets.append('apellido=?'); params.append((d['apellido'] or '').strip() or None)
    if 'rol_predeterminado' in d:
        rol = (d['rol_predeterminado'] or 'todero').strip().lower()
        if rol not in ('dispensacion','envasado','acondicionamiento','todero','jefe'):
            rol = 'todero'
        sets.append('rol_predeterminado=?'); params.append(rol)
    if 'fija_en_dispensacion' in d:
        sets.append('fija_en_dispensacion=?'); params.append(1 if d['fija_en_dispensacion'] else 0)
    if 'es_jefe_produccion' in d:
        sets.append('es_jefe_produccion=?'); params.append(1 if d['es_jefe_produccion'] else 0)
    if 'activo' in d:
        sets.append('activo=?'); params.append(1 if d['activo'] else 0)
    if not sets:
        return jsonify({'error': 'nada que actualizar'}), 400
    params.append(op_id)
    cur = c.execute(f"UPDATE operarios_planta SET {', '.join(sets)} WHERE id=?", params)
    conn.commit()
    if cur.rowcount == 0:
        return jsonify({'error': 'no encontrado'}), 404
    return jsonify({'ok': True, 'id': op_id})


@bp.route('/api/programacion/programar/<int:evento_id>/asignar', methods=['PATCH'])
def prog_asignar_sala_operarios(evento_id):
    """Asigna sala fisica y operarios por fase a una produccion programada.

    Body JSON (todo opcional, NULL para desasignar):
        area_id: int
        operario_dispensacion_id: int
        operario_elaboracion_id: int
        operario_envasado_id: int
        operario_acondicionamiento_id: int

    Valida que:
    - La sala exista y este activa.
    - Si se asigna sala que NO produce y la produccion implica producir → 400.
    - Conflicto: si OTRA produccion ya tomo esa sala ese dia → warning en
      la respuesta (no bloquea — Sebastian decide).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db(); c = conn.cursor()

    # Verificar que la produccion existe
    pp = c.execute(
        "SELECT id, producto, fecha_programada, area_id FROM produccion_programada WHERE id=?",
        (evento_id,)
    ).fetchone()
    if not pp:
        return jsonify({'error': 'produccion no encontrada'}), 404

    updates = []
    params = []
    warnings = []

    # Sala
    if 'area_id' in data:
        new_area = data['area_id']
        if new_area is not None:
            try:
                new_area = int(new_area)
            except (TypeError, ValueError):
                return jsonify({'error': 'area_id invalido'}), 400
            sala = c.execute(
                "SELECT id, nombre, puede_producir FROM areas_planta WHERE id=? AND activo=1",
                (new_area,)
            ).fetchone()
            if not sala:
                return jsonify({'error': 'sala no existe o inactiva'}), 400
            # Conflicto: misma fecha, sala ya tomada por otra produccion
            conflictos = c.execute("""
                SELECT id, producto FROM produccion_programada
                WHERE fecha_programada=? AND area_id=? AND id<>?
                  AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
            """, (pp[2], new_area, evento_id)).fetchall()
            if conflictos:
                warnings.append({
                    'tipo': 'sala_ocupada',
                    'sala': sala[1],
                    'fecha': pp[2],
                    'choca_con': [{'id': x[0], 'producto': x[1]} for x in conflictos],
                })
        updates.append('area_id=?'); params.append(new_area)

    # Operarios por fase
    for campo in ('operario_dispensacion_id', 'operario_elaboracion_id',
                  'operario_envasado_id', 'operario_acondicionamiento_id'):
        if campo in data:
            val = data[campo]
            if val is not None:
                try:
                    val = int(val)
                except (TypeError, ValueError):
                    return jsonify({'error': f'{campo} invalido'}), 400
                op = c.execute(
                    "SELECT id, nombre FROM operarios_planta WHERE id=? AND activo=1",
                    (val,)
                ).fetchone()
                if not op:
                    return jsonify({'error': f'{campo}: operario no existe'}), 400
            updates.append(f'{campo}=?'); params.append(val)

    if not updates:
        return jsonify({'error': 'nada que actualizar'}), 400

    params.append(evento_id)
    c.execute(
        f"UPDATE produccion_programada SET {', '.join(updates)} WHERE id=?",
        params
    )
    # Sebastián 25-may-2026 · audit zero-error · operación mutante en
    # produccion_programada (asignar sala/operarios) requiere audit_log
    # según CLAUDE.md MEMORY (mandatory en cualquier UPDATE de la tabla).
    # Antes esta función UPDATEaba sin dejar rastro · pérdida de trazabilidad
    # ante disputas de asignación operario.
    try:
        from audit_helpers import audit_log
        _data_snap = {k: data[k] for k in (
            'area_id', 'operario_dispensacion_id', 'operario_elaboracion_id',
            'operario_envasado_id', 'operario_acondicionamiento_id'
        ) if k in data}
        audit_log(c, usuario=session.get('compras_user', ''),
                  accion='ASIGNAR_SALA_OPERARIOS',
                  tabla='produccion_programada', registro_id=str(evento_id),
                  antes={'area_id_anterior': pp[3] if len(pp) > 3 else None,
                          'producto': pp[1], 'fecha': pp[2]},
                  despues=_data_snap,
                  detalle=f"Asignación pp {evento_id} ({pp[1]}) · {len(updates)} campos")
    except Exception as _e_audit:
        # Falla silenciosa solo en audit · no romper el flujo principal
        import logging as _lg
        _lg.getLogger('programacion').warning(
            'audit_log ASIGNAR_SALA_OPERARIOS fallo: %s', _e_audit)
    conn.commit()
    return jsonify({'ok': True, 'id': evento_id, 'warnings': warnings})


@bp.route('/api/planta/operarios/historial', methods=['GET'])
def planta_historial_operarios():
    """Devuelve cuantos dias lleva cada operario en cada fase, ultimos 14 dias.
    Sirve para sugerir rotacion: si Camilo lleva 5 dias en acondicionamiento,
    el UI lo marca con un badge 'rotar'. Mayerlin se excluye (es fija)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    # Cuenta apariciones por operario+fase en producciones de los ultimos 14 dias
    sql = """
        SELECT op.id, op.nombre, op.apellido, fase, COUNT(*) as veces
        FROM operarios_planta op
        LEFT JOIN (
            SELECT operario_dispensacion_id as op_id, 'dispensacion' as fase, fecha_programada
              FROM produccion_programada
              WHERE operario_dispensacion_id IS NOT NULL
                AND fecha_programada >= date('now', '-5 hours', '-14 day')
                AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado')
            UNION ALL
            SELECT operario_elaboracion_id, 'elaboracion', fecha_programada
              FROM produccion_programada
              WHERE operario_elaboracion_id IS NOT NULL
                AND fecha_programada >= date('now', '-5 hours', '-14 day')
                AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado')
            UNION ALL
            SELECT operario_envasado_id, 'envasado', fecha_programada
              FROM produccion_programada
              WHERE operario_envasado_id IS NOT NULL
                AND fecha_programada >= date('now', '-5 hours', '-14 day')
                AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado')
            UNION ALL
            SELECT operario_acondicionamiento_id, 'acondicionamiento', fecha_programada
              FROM produccion_programada
              WHERE operario_acondicionamiento_id IS NOT NULL
                AND fecha_programada >= date('now', '-5 hours', '-14 day')
                AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado')
        ) hist ON hist.op_id = op.id
        WHERE op.activo=1 AND op.es_jefe_produccion=0
        GROUP BY op.id, fase
        ORDER BY op.nombre, fase
    """
    raw = conn.execute(sql).fetchall()
    # Reformatear: { op_id: {nombre, fija, fases: {fase: count}} }
    result = {}
    for r in raw:
        op_id, nombre, apellido, fase, veces = r
        if op_id not in result:
            result[op_id] = {
                'id': op_id, 'nombre': nombre, 'apellido': apellido or '',
                'fases': {}, 'total': 0,
            }
        if fase:
            result[op_id]['fases'][fase] = veces
            result[op_id]['total'] += veces
    # Marcar quien necesita rotar (>=4 veces seguidas en una sola fase)
    for op in result.values():
        for fase, cnt in op['fases'].items():
            if cnt >= 4 and op['total'] == cnt:
                op['sugerir_rotar'] = True
                op['fase_acumulada'] = fase
                op['dias_en_fase'] = cnt
                break
    return jsonify({'operarios': list(result.values()), 'ventana_dias': 14})


@bp.route('/api/programacion/programar/<int:evento_id>/iniciar', methods=['POST'])
def prog_iniciar_produccion(evento_id):
    """Operario aprieta 'Iniciar produccion' — graba inicio_real_at,
    marca la sala asignada como 'ocupada' y DESCUENTA INVENTARIO MP.

    Sebastian 5-may-2026 (Luis Enrique): "cuando carguen produccion de
    materia prima de una descuente". Antes el descuento estaba en
    /completar pero las producciones quedaban en envasado y nunca se
    completaban → MPs no salian del inventario. Ahora el descuento es
    al INICIAR (cuando el operario fisicamente saca MP de bodega).

    Comportamiento:
      1. Atomic claim de inventario_descontado_at (idempotencia race-safe)
      2. Calcula MPs desde formula_items (cantidad_g_por_lote * lotes,
         fallback a porcentaje * kg)
      3. Pre-check stock: si falta para alguna MP → 422 sin tocar nada
      4. INSERT movimientos tipo 'Salida' por lote (FEFO)
      5. Set inicio_real_at + sala='ocupada'
      6. Audit log INICIAR_PRODUCCION con detalle de MPs descontadas

    Si descuento falla → rollback completo, inicio_real_at NO se setea
    (la producción queda en estado 'pendiente_iniciar' lista para
    re-intentar tras corregir stock).

    Idempotente: si ya tiene inicio_real_at, retorna ok=True con el ts
    existente.

    Body opcional: {forzar_redescuento: true} para admin (re-iniciar
    producción cuyo descuento fue revertido manualmente).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    body = request.get_json(silent=True) or {}
    forzar = bool(body.get('forzar_redescuento'))
    if forzar and user not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin puede forzar re-descuento'}), 403

    conn = get_db(); c = conn.cursor()
    # BUG-1 fix · 19-may-2026: validar que el caller esté asignado a esta
    # producción (o sea admin/jefe). Antes solo se chequeaba login → un
    # operario podía iniciar/descontar MPs de producción ajena.
    # P0-2 · 23-may-PM · auditoría agente · /iniciar descuenta MP ·
    # solo el operario de dispensación (o jefe/admin) puede.
    ok_caller, err_caller = _caller_puede_operar_produccion(
        c, user, evento_id, rol_requerido='dispensacion')
    if not ok_caller:
        return err_caller
    pp = c.execute("""SELECT id, producto, area_id, inicio_real_at, fin_real_at,
                              COALESCE(inventario_descontado_at,'') as desc_at
                      FROM produccion_programada WHERE id=?""", (evento_id,)).fetchone()
    if not pp:
        return jsonify({'error': 'produccion no existe'}), 404
    if pp[3]:  # ya iniciada
        return jsonify({
            'ok': True, 'ya_iniciada': True, 'inicio_real_at': pp[3],
            'inventario_descontado_at': pp[5] or None,
        })
    if pp[4]:  # ya terminada
        return jsonify({'error': 'produccion ya terminada'}), 400

    # BUG-5 fix · 20-may-2026 OLA 1: GATE SALA SUCIA. UI Mi Día solo MUESTRA
    # un confirm("¿iniciar de todos modos?") pero el backend aceptaba y
    # descontaba MP en sala sucia → contaminación cruzada post-INVIMA.
    # Ahora backend BLOQUEA · admin puede saltarlo con ?force_sucia=1
    # (registrado en audit_log con motivo).
    area_id = pp[2]
    force_sucia = bool(body.get('force_sucia'))
    if area_id and not force_sucia:
        area_row = c.execute(
            "SELECT estado, codigo FROM areas_planta WHERE id=? AND activo=1",
            (area_id,),
        ).fetchone()
        if area_row and (area_row[0] or '').lower() == 'sucia':
            return jsonify({
                'ok': False,
                'error': (f'Sala {area_row[1]} está SUCIA · debe marcarse '
                          'limpia antes de iniciar producción (contaminación '
                          'cruzada · BPM/INVIMA). Admin puede saltarlo con '
                          'force_sucia=true + motivo en observaciones.'),
                'codigo': 'SALA_SUCIA',
                'sala_codigo': area_row[1],
            }), 409
    if area_id and force_sucia and user not in ADMIN_USERS:
        return jsonify({
            'ok': False,
            'error': 'Solo admin puede saltarse el gate de sala sucia',
        }), 403
    if force_sucia:
        log.warning(
            'INICIAR_PRODUCCION OVERRIDE sala sucia · user=%s evento=%s',
            user, evento_id)

    # ── DESCONTAR INVENTARIO MP ATOMICAMENTE ──────────────────────────
    descuento = None
    try:
        descuento = _descontar_mp_produccion(c, evento_id, user, forzar=forzar)
    except _DescuentoError as e:
        try: conn.rollback()
        except Exception: pass
        codigo_http = {
            'SIN_STOCK': 422,
            'SIN_FORMULA': 422,
            'YA_DESCONTADO': 409,
            'PRODUCCION_NO_EXISTE': 404,
        }.get(e.codigo, 500)
        log.warning('iniciar_produccion bloqueado · codigo=%s evento_id=%s msg=%s',
                     e.codigo, evento_id, e)
        return jsonify({
            'ok': False,
            'error': str(e),
            'codigo': e.codigo,
            **e.payload,
        }), codigo_http

    # ── REGISTRAR INICIO + SALA OCUPADA ───────────────────────────────
    c.execute("UPDATE produccion_programada SET inicio_real_at=datetime('now', '-5 hours') WHERE id=?",
              (evento_id,))
    sin_formula = bool(descuento.get('sin_formula'))
    nota_descuento = ('SIN FORMULA · sin descuento' if sin_formula
                       else f'MP descontada {descuento["total_g"]:.0f}g')
    if pp[2]:  # tiene sala asignada → marcar ocupada
        prev = c.execute("SELECT estado FROM areas_planta WHERE id=?", (pp[2],)).fetchone()
        c.execute("UPDATE areas_planta SET estado='ocupada' WHERE id=?", (pp[2],))
        c.execute("""INSERT INTO area_eventos
            (area_id, tipo, estado_anterior, estado_nuevo, produccion_id, usuario, nota)
            VALUES (?,?,?,?,?,?,?)""",
            (pp[2], 'iniciar_prod', (prev[0] if prev else None), 'ocupada',
             evento_id, user, f'inicio: {pp[1]} · {nota_descuento}'))
    audit_log(
        c, usuario=user, accion='INICIAR_PRODUCCION',
        tabla='produccion_programada', registro_id=evento_id,
        despues={
            'producto': pp[1], 'area_id': pp[2],
            'inventario_descontado_at': descuento['inventario_descontado_at'],
            'mps_descontadas_count': len(descuento['mps_descontadas']),
            'total_g_descontado': descuento['total_g'],
            'sin_formula': sin_formula,
            'forzar': forzar,
        },
        detalle=(f"Inició producción {pp[1]} (id={evento_id}) · " +
                  (f"SIN formula valida — NO se descontó inventario"
                   if sin_formula
                   else f"descontó {len(descuento['mps_descontadas'])} MPs "
                        f"({descuento['total_g']:.0f}g)")),
    )

    # ── BRD auto-EBR · si hay MBR aprobado, crear EBR vinculado ───────
    # NON-FATAL: si falla la creación del EBR, NO se aborta el inicio de
    # producción. El flujo crítico (Mayerlin/operario inicia lote) NO debe
    # bloquearse por culpa del BRD. Loguear warning + retornar info en
    # response para visibilidad.
    ebr_info = _intentar_crear_ebr_auto(c, evento_id, pp[1],
                                        descuento.get('total_g'), user)

    conn.commit()
    return jsonify({
        'ok': True,
        'evento_id': evento_id,
        'inicio_real_at': c.execute(
            "SELECT inicio_real_at FROM produccion_programada WHERE id=?",
            (evento_id,)).fetchone()[0],
        'inventario_descontado_at': descuento['inventario_descontado_at'],
        'mps_descontadas': descuento['mps_descontadas'],
        'total_g_descontado': descuento['total_g'],
        'sin_formula': sin_formula,
        'warning': descuento.get('warning'),
        'brd_ebr': ebr_info,
    })


def _intentar_crear_ebr_auto(c, evento_id, producto, total_g_descontado, user):
    """Crea EBR vinculado a una producción cuando se inicia desde Calendar.

    Retorna dict con resultado:
      - ok=True + ebr_id si se creó
      - ok=False + razon si no se creó (sin MBR aprobado, ya existe, error)

    NON-FATAL: ningún error propaga. Cualquier excepción se atrapa y se
    loguea para no romper el flujo de iniciar producción.
    """
    try:
        mbr = c.execute(
            """SELECT id, version, lote_size_g
               FROM mbr_templates
               WHERE producto_nombre = ? AND estado = 'aprobado'
               ORDER BY version DESC LIMIT 1""",
            (producto,),
        ).fetchone()
        if not mbr:
            return {'ok': False, 'razon': 'sin MBR aprobado para este producto',
                     'producto': producto}

        # Idempotencia: si ya existe EBR para esta producción, no duplicar
        existe = c.execute(
            "SELECT id, lote FROM ebr_ejecuciones WHERE produccion_id = ?",
            (evento_id,),
        ).fetchone()
        if existe:
            return {'ok': True, 'ya_existia': True, 'ebr_id': existe[0],
                     'lote': existe[1]}

        # Generar lote único basado en evento_id + fecha
        from datetime import datetime as _dt, timezone as _tz
        fecha = _dt.now(_tz.utc).strftime('%Y%m%d')
        # producto_corto: primeras 3 palabras → siglas + sin espacios
        palabras = (producto or 'PROD').split()[:3]
        prod_short = ''.join(p[:3].upper() for p in palabras)[:12]
        lote = f"{prod_short}-{evento_id}-{fecha}"

        # cantidad_objetivo_g: usar total descontado si existe (más preciso),
        # o lote_size_g del MBR como fallback
        cantidad_obj = total_g_descontado or mbr[2]

        # numero_op MyBatch-compat (mig 117) · atómico via op_counters
        from blueprints.brd import assign_numero_op
        numero_op = assign_numero_op(c)

        # BUG-8 fix · 19-may-2026 audit Planta PERFECTA: SAVEPOINT antes de
        # crear ebr_ejecuciones y clonar pasos · si el clonado falla a mitad
        # de bucle, ROLLBACK TO SAVEPOINT deja la BD consistente (NO crea
        # ebr_ejecuciones huérfano con pasos parciales). El SAVEPOINT vive
        # dentro de la transacción del caller · solo aborta el bloque EBR.
        c.execute('SAVEPOINT ebr_auto')
        try:
            c.execute(
                """INSERT INTO ebr_ejecuciones
                     (mbr_template_id, mbr_version, produccion_id, lote, numero_op,
                      estado, iniciado_por, iniciado_at_utc, cantidad_objetivo_g, notas)
                   VALUES (?, ?, ?, ?, ?, 'iniciado', ?, datetime('now', '-5 hours', 'utc'), ?, ?)""",
                (mbr[0], mbr[1], evento_id, lote, numero_op, user, cantidad_obj,
                 f'Auto-creado al iniciar producción Calendar id={evento_id}'),
            )
            ebr_id = c.lastrowid

            # Clonar pasos del MBR
            pasos_mbr = c.execute(
                """SELECT id, orden, descripcion, tipo_paso, equipo_requerido,
                          requiere_e_sign, requiere_qc
                   FROM mbr_pasos WHERE mbr_template_id = ? ORDER BY orden""",
                (mbr[0],),
            ).fetchall()
            for p in pasos_mbr:
                c.execute(
                    """INSERT INTO ebr_pasos_ejecutados
                         (ebr_id, mbr_paso_id, orden, descripcion, tipo_paso,
                          equipo_requerido, requiere_e_sign, requiere_qc, estado)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pendiente')""",
                    (ebr_id, p[0], p[1], p[2], p[3], p[4], p[5], p[6]),
                )
            c.execute('RELEASE SAVEPOINT ebr_auto')
        except Exception as _ebr_e:
            try:
                c.execute('ROLLBACK TO SAVEPOINT ebr_auto')
                c.execute('RELEASE SAVEPOINT ebr_auto')
            except Exception:
                pass
            raise _ebr_e

        log.info('BRD auto-EBR creado · evento=%s ebr=%s lote=%s op=%s pasos=%d',
                  evento_id, ebr_id, lote, numero_op, len(pasos_mbr))
        return {'ok': True, 'ebr_id': ebr_id, 'lote': lote,
                 'numero_op': numero_op,
                 'pasos_clonados': len(pasos_mbr),
                 'mbr_version': mbr[1]}

    except Exception as e:
        log.warning('BRD auto-EBR falló (NO bloquea inicio) · evento=%s err=%s',
                     evento_id, e)
        return {'ok': False, 'razon': f'excepción: {str(e)[:200]}',
                 'producto': producto}


@bp.route('/api/programacion/programar/<int:evento_id>/terminar', methods=['POST'])
def prog_terminar_produccion(evento_id):
    """Operario aprieta 'Terminar' — graba fin_real_at, marca sala 'sucia'
    (asume que despues de produccion siempre hay limpieza), registra evento.

    Sebastián 12-may-2026 (Planta Mejora A): acepta body opcional
    {kg_real, unidades_real, observaciones} para capturar cantidad real
    producida y calcular merma_pct = (1 - kg_real/cantidad_kg) * 100.
    El operario puede omitirlo (queda pendiente · UI lo recordará).

    Idempotente: si ya tiene fin_real_at, no sobreescribe."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    body = request.get_json(silent=True) or {}
    conn = get_db(); c = conn.cursor()
    # BUG-1 fix · 19-may-2026: solo el operario asignado / jefe / admin
    # puede terminar esta producción.
    ok_caller, err_caller = _caller_puede_operar_produccion(c, user, evento_id)
    if not ok_caller:
        return err_caller
    pp = c.execute("""SELECT id, producto, area_id, inicio_real_at, fin_real_at,
                              cantidad_kg
                      FROM produccion_programada WHERE id=?""", (evento_id,)).fetchone()
    if not pp:
        return jsonify({'error': 'produccion no existe'}), 404
    if pp[4]:
        return jsonify({'ok': True, 'ya_terminada': True, 'fin_real_at': pp[4]})
    if not pp[3]:
        return jsonify({'error': 'produccion no ha iniciado'}), 400

    # Validar kg_real / unidades_real si se reportan
    kg_real = body.get('kg_real')
    unidades_real = body.get('unidades_real')
    merma_pct = None
    cantidad_kg = pp[5] or 0
    if kg_real is not None:
        try:
            kg_real = float(kg_real)
        except (TypeError, ValueError):
            return jsonify({'error': 'kg_real inválido'}), 400
        if kg_real < 0:
            return jsonify({'error': 'kg_real debe ser >= 0'}), 400
        # Sanity check 70-110% del planeado · fuera de eso, requiere
        # observaciones explícitas para auditoría.
        if cantidad_kg > 0:
            ratio = kg_real / cantidad_kg
            if (ratio < 0.7 or ratio > 1.1) and not (body.get('observaciones') or '').strip():
                return jsonify({
                    'error': f'kg_real fuera de rango 70-110% del planeado ({cantidad_kg} kg). '
                              f'Reportar observaciones obligatorio.',
                    'planeado_kg': cantidad_kg, 'real_kg': kg_real,
                    'ratio_pct': round(ratio * 100, 2),
                }), 400
            # BUG-7 fix · 19-may-2026 audit Planta PERFECTA · INVIMA distingue
            # merma (pérdida real, kg_real<planeado) de sobreproducción
            # (kg_real>planeado). Si reportamos merma_pct=-10% para una
            # sobreproducción del 10%, los reportes a INVIMA confunden ambos
            # casos y la auditoría queda contaminada. Clamp merma a >=0:
            # cuando hay sobreproducción, merma_pct=0 (no se "ganó" MP) y
            # la diferencia queda visible en kg_real vs cantidad_kg.
            merma_pct = round(max((1 - ratio) * 100, 0.0), 2)
    if unidades_real is not None:
        try:
            unidades_real = int(unidades_real)
        except (TypeError, ValueError):
            return jsonify({'error': 'unidades_real inválido'}), 400

    c.execute("""UPDATE produccion_programada
                   SET fin_real_at=datetime('now', '-5 hours'),
                       kg_real=COALESCE(?, kg_real),
                       unidades_real=COALESCE(?, unidades_real),
                       merma_pct=COALESCE(?, merma_pct)
                 WHERE id=?""",
              (kg_real, unidades_real, merma_pct, evento_id))
    if pp[2]:
        prev = c.execute("SELECT estado FROM areas_planta WHERE id=?", (pp[2],)).fetchone()
        # Sebastián 1-may-2026: guard contra sobrescribir 'limpiando'
        c.execute("""
            UPDATE areas_planta SET estado='sucia'
            WHERE id=? AND estado IN ('ocupada','libre')
        """, (pp[2],))
        c.execute("""INSERT INTO area_eventos
            (area_id, tipo, estado_anterior, estado_nuevo, produccion_id, usuario, nota)
            VALUES (?,?,?,?,?,?,?)""",
            (pp[2], 'terminar_prod', (prev[0] if prev else None), 'sucia',
             evento_id, user, f'fin: {pp[1]} → sala sucia, espera limpieza'))
    # Calcular cycle time real
    cycle = c.execute("""SELECT
        CAST((julianday(fin_real_at) - julianday(inicio_real_at)) * 24 * 60 AS INTEGER) as min
        FROM produccion_programada WHERE id=?""", (evento_id,)).fetchone()
    cycle_min = cycle[0] if cycle else None
    audit_log(c, usuario=user, accion='TERMINAR_PRODUCCION', tabla='produccion_programada',
              registro_id=evento_id,
              despues={'producto': pp[1], 'area_id': pp[2], 'cycle_time_min': cycle_min,
                        'kg_real': kg_real, 'unidades_real': unidades_real,
                        'merma_pct': merma_pct},
              detalle=f"Terminó producción {pp[1]} (id={evento_id})"
                      + (f" · cycle {cycle_min} min" if cycle_min else "")
                      + (f" · {kg_real} kg real" if kg_real is not None else "")
                      + (f" · merma {merma_pct}%" if merma_pct is not None else ""))
    conn.commit()
    return jsonify({'ok': True, 'evento_id': evento_id,
                    'cycle_time_min': cycle_min,
                    'kg_real': kg_real,
                    'unidades_real': unidades_real,
                    'merma_pct': merma_pct,
                    'merma_alta': bool(merma_pct is not None and abs(merma_pct) > 5.0)})


@bp.route('/api/planta/listo-producir/<path:producto>', methods=['GET'])
def planta_listo_producir(producto):
    """Semaforo de insumos para una produccion programada.

    Sebastian (30-abr-2026): "respuesta inmediata" — quiere ver de un
    vistazo si una produccion tiene todos los MPs disponibles antes
    de programarla. Antes era email "URGENTE confirmacion insumos".

    Query params:
        lotes: int (default 1) — para calcular requerimiento total
                (cantidad_mp = porcentaje * lote_size_kg * lotes * 1000g/kg)

    Devuelve para cada material:
        codigo_mp, nombre, requerido_g, disponible_g, status
        (✅ ok / ⚠ justo / ❌ deficit)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        lotes = int(request.args.get('lotes', 1))
    except (TypeError, ValueError):
        lotes = 1
    conn = get_db(); c = conn.cursor()

    # Buscar formula y lote_size
    fh = c.execute(
        "SELECT producto_nombre, lote_size_kg, unidad_base_g "
        "FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    if not fh:
        return jsonify({'error': 'producto no tiene formula registrada',
                        'items': []}), 404
    nombre_fh, lote_kg, unidad_g = fh
    # cantidad total en g = lote_kg * 1000 * lotes
    total_g = (lote_kg or 0) * 1000 * max(1, lotes)

    items = c.execute("""
        SELECT fi.material_id, fi.material_nombre, fi.porcentaje,
               COALESCE(fi.cantidad_g_por_lote, 0) as cant_lote_g
        FROM formula_items fi
        WHERE UPPER(TRIM(fi.producto_nombre))=UPPER(TRIM(?))
        ORDER BY fi.porcentaje DESC
    """, (nombre_fh,)).fetchall()

    out = []
    deficit_count = 0; justo_count = 0
    for mat_id, mat_nom, pct, cant_lote_g in items:
        # requerido_g: preferir cantidad_g_por_lote si está, si no calcular % * total
        if cant_lote_g and cant_lote_g > 0:
            req_g = cant_lote_g * lotes
        else:
            req_g = (pct or 0) / 100.0 * total_g

        # Disponible en bodega · excluye CUARENTENA/VENCIDO/RECHAZADO
        # Audit zero-error 2-may-2026: SQL anterior usaba 'Ingreso'/'Consumo'
        # (no existen) y devolvía valores negativos siempre · semáforo roto.
        # FIX 1-jun-2026 · resolver id de fórmula → id de bodega (caso glucosamina)
        # antes del lookup · el semáforo daba 'falta' con stock real.
        disp_g = stock_mp_disponible(c, _resolver_material_bodega(c, mat_id, mat_nom))

        # Status
        if disp_g >= req_g:
            status = 'ok'
        elif disp_g >= req_g * 0.5:
            status = 'justo'; justo_count += 1
        else:
            status = 'deficit'; deficit_count += 1

        out.append({
            'codigo_mp': mat_id,
            'nombre': mat_nom,
            'porcentaje': pct,
            'requerido_g': round(req_g),
            'disponible_g': round(disp_g),
            'status': status,
            'faltante_g': max(0, round(req_g - disp_g)),
        })

    return jsonify({
        'producto': nombre_fh,
        'lote_size_kg': lote_kg,
        'lotes': lotes,
        'total_g': round(total_g),
        'items': out,
        'resumen': {
            'total': len(out),
            'ok': len(out) - justo_count - deficit_count,
            'justo': justo_count,
            'deficit': deficit_count,
            'listo': deficit_count == 0,
        }
    })


# ─── Actividades por operario en sala (turnos con timer) ──────────────────
@bp.route('/api/planta/areas/<int:area_id>/actividades', methods=['GET', 'POST'])
def planta_actividades_sala(area_id):
    """GET: lista actividades de la sala (filtro ?activas=1 → solo en curso).
    POST: inicia un turno de un operario en esta sala.
       body: operario_id (req), tipo (req), descripcion (opt), produccion_id (opt)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()

    if request.method == 'POST':
        d = request.get_json(force=True, silent=True) or {}
        try:
            op_id = int(d.get('operario_id') or 0)
        except (TypeError, ValueError):
            op_id = 0
        if not op_id:
            return jsonify({'error': 'operario_id requerido'}), 400
        tipo = (d.get('tipo') or 'produccion').strip().lower()
        if tipo not in ('produccion','dispensacion','envasado','acondicionamiento',
                        'conteo_ciclico','limpieza','mantenimiento','otro'):
            tipo = 'otro'
        # Verificar operario existe
        op_row = c.execute("SELECT nombre, apellido FROM operarios_planta WHERE id=? AND activo=1",
                           (op_id,)).fetchone()
        if not op_row:
            return jsonify({'error': 'operario no existe o inactivo'}), 400
        # Si ya hay un turno activo del mismo operario en cualquier sala, lo cerramos primero
        # (un operario no puede estar en 2 salas a la vez)
        previo = c.execute("""SELECT id, area_id FROM actividades_sala
                              WHERE operario_id=? AND fin_at IS NULL""",
                           (op_id,)).fetchone()
        cerrado_previo = None
        if previo:
            prev_id, prev_area = previo
            c.execute("""UPDATE actividades_sala
                SET fin_at=datetime('now', '-5 hours'),
                    duracion_min=CAST((julianday(datetime('now', '-5 hours'))-julianday(inicio_at))*24*60 AS INTEGER)
                WHERE id=?""", (prev_id,))
            cerrado_previo = {'actividad_id': prev_id, 'area_id': prev_area}
        # Insertar actividad nueva
        c.execute("""INSERT INTO actividades_sala
            (area_id, operario_id, tipo, descripcion, produccion_id, creado_por)
            VALUES (?,?,?,?,?,?)""",
            (area_id, op_id, tipo,
             (d.get('descripcion') or '').strip() or None,
             d.get('produccion_id') or None, user))
        new_id = c.lastrowid
        # Cambiar estado de la sala a 'ocupada' si no lo estaba
        c.execute("UPDATE areas_planta SET estado='ocupada' WHERE id=? AND estado='libre'", (area_id,))
        # Log evento
        try:
            c.execute("""INSERT INTO area_eventos
                (area_id, tipo, estado_anterior, estado_nuevo, produccion_id, operario_id, usuario, nota)
                VALUES (?,?,?,?,?,?,?,?)""",
                (area_id, 'iniciar_prod', 'libre', 'ocupada',
                 d.get('produccion_id'), op_id, user,
                 f'turno {tipo} de {op_row[0]}'))
        except Exception:
            pass
        conn.commit()
        return jsonify({
            'ok': True,
            'actividad_id': new_id,
            'cerrado_previo': cerrado_previo,
            'operario': (op_row[0] + ' ' + (op_row[1] or '')).strip(),
        }), 201

    # GET — actividades de la sala
    solo_activas = request.args.get('activas') == '1'
    sql = """SELECT a.id, a.operario_id, a.tipo, a.descripcion, a.produccion_id,
                    a.inicio_at, a.fin_at, a.duracion_min, a.observaciones,
                    op.nombre, op.apellido,
                    pp.producto,
                    CAST((julianday('now')-julianday(a.inicio_at))*24*60 AS INTEGER) as min_corridos
             FROM actividades_sala a
             LEFT JOIN operarios_planta op ON op.id = a.operario_id
             LEFT JOIN produccion_programada pp ON pp.id = a.produccion_id
             WHERE a.area_id=?"""
    params = [area_id]
    if solo_activas:
        sql += " AND a.fin_at IS NULL"
    sql += " ORDER BY a.inicio_at DESC LIMIT 50"
    rows = c.execute(sql, params).fetchall()
    out = []
    for r in rows:
        out.append({
            'id': r[0], 'operario_id': r[1], 'tipo': r[2],
            'descripcion': r[3], 'produccion_id': r[4],
            'inicio_at': r[5], 'fin_at': r[6],
            'duracion_min': r[7] if r[6] else r[12],
            'en_curso': r[6] is None,
            'observaciones': r[8],
            'operario_nombre': ((r[9] or '') + ' ' + (r[10] or '')).strip(),
            'producto': r[11],
            'minutos_corridos': r[12] if not r[6] else None,
        })
    return jsonify({'actividades': out, 'area_id': area_id})


@bp.route('/api/planta/actividades/<int:act_id>/terminar', methods=['POST'])
def planta_actividad_terminar(act_id):
    """Cierra el turno de un operario. Calcula duracion_min."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(force=True, silent=True) or {}
    obs = (d.get('observaciones') or '').strip() or None
    conn = get_db(); c = conn.cursor()
    row = c.execute("""SELECT id, area_id, operario_id, fin_at FROM actividades_sala
                       WHERE id=?""", (act_id,)).fetchone()
    if not row:
        return jsonify({'error': 'actividad no existe'}), 404
    if row[3]:
        return jsonify({'ok': True, 'ya_cerrada': True})
    c.execute("""UPDATE actividades_sala
        SET fin_at=datetime('now', '-5 hours'),
            duracion_min=CAST((julianday(datetime('now', '-5 hours'))-julianday(inicio_at))*24*60 AS INTEGER),
            observaciones=COALESCE(?, observaciones)
        WHERE id=?""", (obs, act_id))
    # Si no quedan actividades activas en la sala → marcar libre o sucia
    pendientes = c.execute(
        "SELECT COUNT(*) FROM actividades_sala WHERE area_id=? AND fin_at IS NULL",
        (row[1],)
    ).fetchone()[0]
    if pendientes == 0:
        # Si la sala estaba ocupada, pasarla a sucia (necesita limpieza)
        c.execute("""UPDATE areas_planta SET estado='sucia'
                     WHERE id=? AND estado='ocupada'""", (row[1],))
        try:
            c.execute("""INSERT INTO area_eventos
                (area_id, tipo, estado_anterior, estado_nuevo, operario_id, usuario, nota)
                VALUES (?,?,?,?,?,?,?)""",
                (row[1], 'terminar_prod', 'ocupada', 'sucia',
                 row[2], user, 'ultimo turno cerrado'))
        except Exception:
            pass
    conn.commit()
    cycle = c.execute("SELECT duracion_min FROM actividades_sala WHERE id=?", (act_id,)).fetchone()
    return jsonify({'ok': True, 'duracion_min': cycle[0] if cycle else None})


@bp.route('/api/planta/actividades/kpis', methods=['GET'])
def planta_actividades_kpis():
    """KPIs agregados de actividades para indicadores. Filtros:
       ?desde=YYYY-MM-DD  ?hasta=YYYY-MM-DD  (default: ultimos 30 dias)"""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    desde = request.args.get('desde') or ''
    hasta = request.args.get('hasta') or ''
    if not desde:
        desde = (date.today() - timedelta(days=30)).isoformat()
    if not hasta:
        hasta = date.today().isoformat()
    conn = get_db()
    # Total min por operario
    rows_op = conn.execute("""
        SELECT op.nombre, op.apellido,
               COUNT(*) as turnos,
               SUM(COALESCE(a.duracion_min, 0)) as min_total
        FROM actividades_sala a
        LEFT JOIN operarios_planta op ON op.id = a.operario_id
        WHERE a.fin_at IS NOT NULL
          AND DATE(a.inicio_at) >= ? AND DATE(a.inicio_at) <= ?
        GROUP BY a.operario_id
        ORDER BY min_total DESC
    """, (desde, hasta)).fetchall()
    por_operario = [{
        'operario': ((r[0] or '') + ' ' + (r[1] or '')).strip(),
        'turnos': r[2] or 0,
        'min_total': r[3] or 0,
        'horas': round((r[3] or 0)/60, 1),
    } for r in rows_op]
    # Total min por tipo
    rows_t = conn.execute("""
        SELECT tipo, COUNT(*) as turnos, SUM(COALESCE(duracion_min,0)) as mins
        FROM actividades_sala
        WHERE fin_at IS NOT NULL
          AND DATE(inicio_at) >= ? AND DATE(inicio_at) <= ?
        GROUP BY tipo
        ORDER BY mins DESC
    """, (desde, hasta)).fetchall()
    por_tipo = [{
        'tipo': r[0], 'turnos': r[1] or 0,
        'horas': round((r[2] or 0)/60, 1),
    } for r in rows_t]
    # Activas ahora
    activas = conn.execute(
        "SELECT COUNT(*) FROM actividades_sala WHERE fin_at IS NULL"
    ).fetchone()[0]
    return jsonify({
        'desde': desde, 'hasta': hasta,
        'por_operario': por_operario,
        'por_tipo': por_tipo,
        'turnos_activos_ahora': activas,
    })


@bp.route('/api/planta/centro-mando', methods=['GET'])
def planta_centro_mando():
    """Endpoint agregado del Centro de Mando — devuelve TODO en una llamada
    para que la UI haga UN fetch y pinte el tablero completo. Optimizado
    para auto-refresh cada 30s.

    Retorna:
      areas: list[{id, codigo, nombre, tipo, estado, capacidades, ocupada_por[]}]
        ocupada_por incluye produccion_id, producto, lotes, kg, operarios y
        inicio_real_at + minutos_corridos para timer en vivo.
      operarios_libres: list — quienes no estan en una produccion en curso.
      kpis: produciones_activas_ahora, prods_terminadas_hoy,
            cycle_time_promedio_min, salas_libres, salas_sucias.
      eventos_recientes: ultimos 10 eventos para timeline lateral.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    # BUG-18 fix · 19-may-2026 audit Planta PERFECTA · restringir a roles
    # de planta/admin · antes cualquier user logueado (contadora, comercial)
    # podía ver KPIs operativos + asignaciones + producto/lotes/kg.
    _user = session.get('compras_user', '')
    _permitidos = set(ADMIN_USERS) | set(COMPRAS_USERS) | set(PLANTA_USERS)
    if _user not in _permitidos:
        return jsonify({'error': 'Solo planta/compras/admin'}), 403
    conn = get_db()
    # Sebastián 1-may-2026: respeta selector de fecha + horizonte 7 días
    # si el día seleccionado está vacío, muestra próximos 7 días (no dejar
    # vacío al jefe cuando hoy es viernes y todo está programado para lunes)
    fecha_param = (request.args.get('fecha') or '').strip()[:10]
    try:
        fecha_sel = datetime.strptime(fecha_param, '%Y-%m-%d').date() if fecha_param else date.today()
    except Exception:
        fecha_sel = date.today()

    # ── AUTO-LIMPIEZA INDEPENDIENTE AL INICIO (Sebastián 1-may-2026) ──
    # 'sigue mostrando lunes todas, martes mezcla' → cancelar huérfanas SQL-first.
    # Cursor explícito + commit forzado + logs de diagnóstico.
    #
    # PERF-FIX · 22-may-2026 · skip cleanup si header X-Skip-Cleanup=1 o
    # ?skip_cleanup=1 · permite que UI pida sin esperar 200 cancelaciones.
    # El cron diario sigue limpiando · este auto-clean inline ahora es opcional.
    auto_canceladas_inicial = 0
    auto_cancel_detalle_inicial = []
    auto_clean_diag = {'cal_set_size': 0, 'db_rows_check': 0, 'matched': 0, 'skipped': False}
    _skip_cleanup = (
        request.headers.get('X-Skip-Cleanup') == '1'
        or request.args.get('skip_cleanup') == '1'
    )
    _ac = conn.cursor()
    if _skip_cleanup:
        auto_clean_diag['skipped'] = True
    try:
        if _skip_cleanup:
            raise StopIteration('skip_cleanup activado')  # salta al except
        _f_inicio = fecha_sel.isoformat()
        _f_fin = (fecha_sel + timedelta(days=6)).isoformat()
        # Set de Calendar (fecha, producto_upper) en horizonte
        _productos_cal = set()
        try:
            from blueprints.auto_plan import (
                _calendar_events_cached as _cec,
                _match_producto_evento as _mpe,
            )
            # PERF-FIX 23-may-2026 · auditoría · era force_refresh=True que
            # bypaseaba cache de 60s · auto-refresh frontend (cada 30s) +
            # cleanup loop sincrono = 2-5s por request · ahora usa cache
            # · cleanup completo lo hace el cron diario auto_reparar_huerfanas
            cal_events = _cec(force_refresh=False) or []
            skus_aliases = {}
            for sku_n, alias_csv in _ac.execute("""
                SELECT producto_nombre, COALESCE(alias_calendar,'')
                FROM sku_planeacion_config
                WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
            """).fetchall():
                skus_aliases[sku_n] = alias_csv
            for ev in cal_events:
                f_ev = (ev.get('fecha') or '')[:10]
                if f_ev < _f_inicio or f_ev > _f_fin: continue
                producto_match = None; best = 0
                for prod_n, alias_csv in skus_aliases.items():
                    try:
                        s = _mpe(prod_n, alias_csv, ev.get('titulo'), ev.get('descripcion',''))
                        if s >= 50 and s > best:
                            best = s; producto_match = prod_n
                    except Exception:
                        continue
                if producto_match:
                    _productos_cal.add((f_ev, producto_match.upper()))
        except Exception as _e:
            logging.getLogger('programacion').warning(f'[centro-mando auto-clean] cal fetch falla: {_e}')
        auto_clean_diag['cal_set_size'] = len(_productos_cal)

        # Sebastián 1-may-2026 estricto Calendar-first: cancelar TODAS las
        # filas DB no iniciadas en horizonte. Calendar es la única fuente
        # de verdad. Si una fila DB no tiene inicio_real_at, es estale o
        # placeholder · debe desaparecer. Si necesita mostrarse, Calendar
        # debe tenerla.
        # Guard único: máximo 200 por GET.
        # Sebastián 19-may-2026 (post-incidente): NUNCA tocar lo FIJO
        # (eos_plan / eos_b2b / eos_retroactivo). El auto-clean estricto
        # Calendar-first cancelaba lo que el usuario fijó si no estaba en
        # Google Calendar · ése fue el bug que borró las 4 producciones de
        # la semana del 19-may. Ahora solo aplica a sugerencias.
        _db_rows = _ac.execute("""
            SELECT id, producto, date(fecha_programada)
            FROM produccion_programada
            WHERE date(fecha_programada) BETWEEN ? AND ?
              AND COALESCE(estado, 'programado') IN ('', 'programado', 'planeado')
              AND inicio_real_at IS NULL
              AND COALESCE(origen,'') NOT IN ('eos_plan','eos_b2b','eos_retroactivo')
            LIMIT 200
        """, (_f_inicio, _f_fin)).fetchall()
        auto_clean_diag['db_rows_check'] = len(_db_rows)
        # P0 audit 26-may · DEFENSIVO · si Calendar set viene VACÍO,
        # NUNCA limpiamos · era el bug que un blip 60s del Google
        # Calendar API wipea el plan semanal entero · re-incidente
        # del 19-may era exactamente este patrón.
        if not _productos_cal:
            auto_clean_diag['skipped_calendar_empty'] = True
            _db_rows = []  # bypass del loop · no cancela nada
        for pid, prod, fecha in _db_rows:
            key = (fecha, (prod or '').upper())
            if key in _productos_cal:
                # Tiene match Calendar exacto → mantener (Calendar la pre-asigna)
                auto_clean_diag['matched'] += 1
                continue
            # No tiene match · cancelar (Calendar set NO está vacío · feed OK)
            _ac.execute("""
                UPDATE produccion_programada
                  SET estado='cancelado',
                      observaciones = COALESCE(observaciones,'') ||
                        ' [auto-cancelado · estricto Calendar-first]'
                WHERE id=?
            """, (pid,))
            auto_canceladas_inicial += 1
            # AUDITORÍA-FIX 23-may-2026 · C16 · liberar SOLs Pre-Prod
            _sols_lib = _liberar_sols_pre_produccion(
                _ac, pid, motivo='Producción auto-cancelada Centro Mando')
            try:
                audit_log(_ac, usuario='sistema_centro_mando_auto_clean',
                          accion='AUTO_CANCELAR_PRODUCCION',
                          tabla='produccion_programada', registro_id=pid,
                          despues={'razon': 'estricto Calendar-first · sin match',
                                   'producto': prod, 'fecha': fecha,
                                   'sols_liberadas': _sols_lib})
            except Exception as _e:
                logging.getLogger('programacion').warning(
                    f'[centro-mando auto-clean] audit fallo id={pid}: {_e}')
            if len(auto_cancel_detalle_inicial) < 5:
                auto_cancel_detalle_inicial.append(f"{(prod or '?')[:30]} {fecha}")
        # FORZAR commit siempre
        conn.commit()
        logging.getLogger('programacion').info(
            f'[centro-mando auto-clean] cal_set={auto_clean_diag["cal_set_size"]} '
            f'db_rows={auto_clean_diag["db_rows_check"]} '
            f'matched={auto_clean_diag["matched"]} '
            f'canceladas={auto_canceladas_inicial}'
        )

        # ── AUTO-RESET SALAS STALE (Sebastián 1-may-2026: 'que quede
        # automático'). Si una sala está 'ocupada' pero NO hay producción
        # en_proceso/iniciado activa en ella → resetear a 'libre'.
        # No tocamos 'sucia' (requiere marcar limpia explícito) ni 'limpiando'.
        salas_reset = 0
        try:
            stale_rows = _ac.execute("""
                SELECT a.id, a.codigo FROM areas_planta a
                WHERE a.estado = 'ocupada'
                  AND a.activo = 1
                  AND NOT EXISTS (
                    SELECT 1 FROM produccion_programada pp
                    WHERE pp.area_id = a.id
                      AND COALESCE(pp.estado,'programado') IN ('en_proceso','iniciado')
                      AND pp.inicio_real_at IS NOT NULL
                      AND pp.fin_real_at IS NULL
                  )
            """).fetchall()
            for area_id, codigo in stale_rows:
                _ac.execute(
                    "UPDATE areas_planta SET estado='libre' WHERE id=? AND estado='ocupada'",
                    (area_id,)
                )
                salas_reset += 1
            if salas_reset:
                conn.commit()
                logging.getLogger('programacion').info(
                    f'[centro-mando auto-reset salas] {salas_reset} salas reset a libre'
                )
        except Exception as _e:
            logging.getLogger('programacion').warning(f'[centro-mando auto-reset salas] falla: {_e}')
        auto_clean_diag['salas_reset'] = salas_reset
    except Exception as _e:
        logging.getLogger('programacion').warning(f'[centro-mando auto-clean inicial] falla: {_e}')
        try: conn.rollback()
        except Exception as _r:
            logging.getLogger('programacion').debug('rollback no aplicable: %s', _r)
    hoy = fecha_sel.isoformat()

    # Areas con producciones activas (in progress: inicio_real_at NOT NULL,
    # fin_real_at NULL) + ultima producción para 'sucia · era X'
    areas = []
    for r in conn.execute("""
        SELECT id, codigo, nombre, tipo, puede_producir, puede_envasar,
               marmita_ml, especial, estado, orden
        FROM areas_planta WHERE activo=1 ORDER BY orden, id
    """):
        areas.append({
            'id': r[0], 'codigo': r[1], 'nombre': r[2], 'tipo': r[3],
            'puede_producir': bool(r[4]), 'puede_envasar': bool(r[5]),
            'marmita_ml': r[6], 'especial': r[7],
            'estado': r[8], 'orden': r[9],
            'ocupada_por': [],
            'ultima_produccion': None,
        })
    area_by_id = {a['id']: a for a in areas}
    # Sebastián 1-may-2026: 'área sucia por fabricar tal'
    # Para cada sala, buscar la ÚLTIMA producción que ocupó/completó esa sala.
    # BUG-20 fix · 19-may-2026 audit Planta PERFECTA: ventana de 30 días y
    # LIMIT 200. Antes el query barría TODA produccion_programada cada 30s
    # (auto-refresh) sin filtro de fecha · degradación lineal con histórico.
    try:
        for r in conn.execute("""
            SELECT pp.area_id, pp.producto, pp.fin_real_at,
                   COALESCE(pp.estado, 'programado') as est
            FROM produccion_programada pp
            WHERE pp.area_id IS NOT NULL
              AND pp.fin_real_at IS NOT NULL
              AND date(pp.fin_real_at) >= date('now', '-5 hours', '-30 day')
            ORDER BY pp.fin_real_at DESC
            LIMIT 200
        """):
            aid = r[0]
            if aid in area_by_id and area_by_id[aid]['ultima_produccion'] is None:
                area_by_id[aid]['ultima_produccion'] = {
                    'producto': r[1], 'terminada_at': r[2], 'estado': r[3],
                }
    except Exception:
        pass
    # Producciones en curso (no terminadas) — vienen siempre, las del dia mas
    # las que se quedaron sin terminar (ocupando sala todavia)
    for r in conn.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes, pp.estado,
               pp.area_id, pp.inicio_real_at, pp.fin_real_at,
               COALESCE(pp.lotes,1)*COALESCE(fh.lote_size_kg,0) as kg,
               od.nombre || ' ' || COALESCE(od.apellido,'')  as op_disp,
               oe.nombre || ' ' || COALESCE(oe.apellido,'')  as op_elab,
               oen.nombre || ' ' || COALESCE(oen.apellido,'') as op_env,
               oa.nombre || ' ' || COALESCE(oa.apellido,'')  as op_acon,
               CAST((julianday('now') - julianday(pp.inicio_real_at)) * 24 * 60 AS INTEGER) as min_corridos
        FROM produccion_programada pp
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre))=UPPER(TRIM(pp.producto))
        LEFT JOIN operarios_planta od  ON od.id  = pp.operario_dispensacion_id
        LEFT JOIN operarios_planta oe  ON oe.id  = pp.operario_elaboracion_id
        LEFT JOIN operarios_planta oen ON oen.id = pp.operario_envasado_id
        LEFT JOIN operarios_planta oa  ON oa.id  = pp.operario_acondicionamiento_id
        WHERE pp.area_id IS NOT NULL
          AND (pp.inicio_real_at IS NOT NULL AND pp.fin_real_at IS NULL
               OR pp.fecha_programada=? AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado'))
        ORDER BY pp.inicio_real_at DESC NULLS LAST
    """, (hoy,)):
        info = {
            'produccion_id': r[0], 'producto': r[1], 'fecha': r[2],
            'lotes': r[3], 'estado': r[4],
            'inicio_real_at': r[6], 'fin_real_at': r[7], 'kg': r[8],
            'minutos_corridos': r[13] if r[6] and not r[7] else None,
            'operario_dispensacion':      (r[9]  or '').strip() or None,
            'operario_elaboracion':       (r[10] or '').strip() or None,
            'operario_envasado':          (r[11] or '').strip() or None,
            'operario_acondicionamiento': (r[12] or '').strip() or None,
            'en_curso': bool(r[6] and not r[7]),
        }
        a = area_by_id.get(r[5])
        if a:
            a['ocupada_por'].append(info)

    # KPIs del dia · CM-FIX #6 · 21-may-2026 · respetar fecha_sel
    # Antes el WHERE usaba date('now','-30 day') siempre · si el user
    # navegaba 30+ días atrás, terminadas_hoy daba 0 confuso. Ahora
    # ventana 30d ANTES de fecha_sel.
    kpi_row = conn.execute("""
        SELECT
            SUM(CASE WHEN inicio_real_at IS NOT NULL AND fin_real_at IS NULL THEN 1 ELSE 0 END) as activas,
            SUM(CASE WHEN DATE(fin_real_at)=? THEN 1 ELSE 0 END) as terminadas_hoy,
            AVG(CASE WHEN DATE(fin_real_at)=?
                     THEN (julianday(fin_real_at)-julianday(inicio_real_at))*24*60
                     ELSE NULL END) as ct_prom_min
        FROM produccion_programada
        WHERE fecha_programada >= date(?, '-30 day')
    """, (hoy, hoy, hoy)).fetchone()
    # BUG-3 fix · 20-may-2026 OLA 1: filtro tipo='produccion' debe aplicarse
    # TAMBIÉN a sucias y ocupadas. Antes sucias y ocupadas contaban ALMP,
    # ALMPT, ACOND, QC, DISP · sumaban estados de áreas que no son de
    # producción · KPIs no cuadraban con realidad ("8 sucias 2 libres"
    # cuando había 0 sucias en producción).
    salas_libres   = sum(1 for a in areas if a['estado']=='libre'   and a['tipo']=='produccion')
    salas_sucias   = sum(1 for a in areas if a['estado']=='sucia'   and a['tipo']=='produccion')
    salas_ocupadas = sum(1 for a in areas if a['estado']=='ocupada' and a['tipo']=='produccion')

    # Eventos recientes (ultimos 15)
    eventos = []
    for r in conn.execute("""
        SELECT ev.tipo, ev.estado_anterior, ev.estado_nuevo,
               ev.usuario, ev.nota, ev.ts,
               ap.codigo, ap.nombre,
               pp.producto
        FROM area_eventos ev
        LEFT JOIN areas_planta ap ON ap.id=ev.area_id
        LEFT JOIN produccion_programada pp ON pp.id=ev.produccion_id
        ORDER BY ev.ts DESC LIMIT 15
    """):
        eventos.append({
            'tipo': r[0], 'de': r[1], 'a': r[2], 'usuario': r[3],
            'nota': r[4], 'ts': r[5], 'area_codigo': r[6],
            'area_nombre': r[7], 'producto': r[8],
        })

    # ── PRODUCCIONES DEL DÍA · Calendar-first (Sebastián 1-may-2026) ──
    # Unifica todo en el mapa: cards arriba muestran lo que toca HOY pero
    # aún no se ha iniciado · click ▶ los marca en proceso en su sala IA.
    # Si el día seleccionado está vacío → muestra próximos 7 días para
    # que el jefe siempre vea qué viene (no quedar en blanco un viernes).
    producciones_dia = []
    db_sin_calendar = []
    capacidad_warnings = []
    fechas_horizonte = [fecha_sel + timedelta(days=i) for i in range(7)]
    fechas_horizonte_iso = set(f.isoformat() for f in fechas_horizonte)

    # Helper IA preview operarios (afinidad ponderada · rotación determinística)
    import hashlib
    op_pool = [(r[0], (r[1] or '').strip(), (r[2] or '').strip().lower())
               for r in conn.execute("""
                    SELECT id, nombre || ' ' || COALESCE(apellido,''),
                           COALESCE(rol_predeterminado,'')
                    FROM operarios_planta
                    WHERE COALESCE(activo,1)=1 AND COALESCE(es_jefe_produccion,0)=0
                    ORDER BY id
               """).fetchall()]
    AFINIDAD_CM = {
        'dispensacion': {'dispensacion': 4, 'elaboracion': 1, 'todero': 2},
        'elaboracion':  {'dispensacion': 3, 'elaboracion': 4, 'todero': 2},
        'envasado':     {'envasado': 4, 'todero': 2},
        'acondicionamiento': {'acondicionamiento': 4, 'envasado': 2, 'todero': 2},
    }
    def _hash_cm(s):
        return int(hashlib.md5(s.encode('utf-8')).hexdigest()[:8], 16)
    def _ops_para(producto_n, fecha_iso):
        if not op_pool: return {}
        roles = ('dispensacion','elaboracion','envasado','acondicionamiento')
        usados = set()
        out = {}
        for rol in roles:
            cands = [(oid, nom, rp) for (oid, nom, rp) in op_pool if oid not in usados]
            if not cands:
                out[rol] = ''
                continue
            afin = AFINIDAD_CM.get(rol, {})
            weighted = [(afin.get(rp, 1) if rp else 1, oid, nom) for (oid, nom, rp) in cands]
            total = sum(w for w, _, _ in weighted)
            target = _hash_cm(f'{producto_n}|{fecha_iso}|{rol}') % max(total, 1)
            cum = 0
            chosen = weighted[0]
            for w, oid, nom in weighted:
                cum += w
                if target < cum:
                    chosen = (w, oid, nom); break
            out[rol] = chosen[2]
            usados.add(chosen[1])
        return out

    try:
        # 1) Filas DB del horizonte SOLO si están INICIADAS o COMPLETADAS.
        # Sebastián 1-may-2026 (estricto Calendar-first): 'se lleva todo lo
        # que hay los lunes sin importar las fechas'. Las filas DB no
        # iniciadas pueden tener fechas erróneas o productos antiguos →
        # ignorarlas completamente. Calendar es la única fuente para 'qué
        # toca'. Solo lo que el operario YA inició (con inicio_real_at) o
        # YA terminó persiste como card.
        rows_db = conn.execute("""
            SELECT pp.id, pp.producto, COALESCE(pp.cantidad_kg,0), pp.lotes,
                   pp.estado, pp.area_id, date(pp.fecha_programada) as f,
                   ap.codigo as area_cod, ap.nombre as area_nom,
                   o1.nombre || ' ' || COALESCE(o1.apellido,'') as op_disp,
                   o2.nombre || ' ' || COALESCE(o2.apellido,'') as op_elab,
                   o3.nombre || ' ' || COALESCE(o3.apellido,'') as op_env,
                   o4.nombre || ' ' || COALESCE(o4.apellido,'') as op_acon,
                   pp.inicio_real_at, pp.fin_real_at
            FROM produccion_programada pp
            LEFT JOIN areas_planta ap ON ap.id = pp.area_id
            LEFT JOIN operarios_planta o1 ON o1.id = pp.operario_dispensacion_id
            LEFT JOIN operarios_planta o2 ON o2.id = pp.operario_elaboracion_id
            LEFT JOIN operarios_planta o3 ON o3.id = pp.operario_envasado_id
            LEFT JOIN operarios_planta o4 ON o4.id = pp.operario_acondicionamiento_id
            WHERE date(pp.fecha_programada) BETWEEN ? AND ?
              AND COALESCE(pp.estado, 'programado') != 'cancelado'
            -- CM-FIX #2 · 20-may-2026 OLA Funcional: SACAR el filtro
            -- inicio_real_at NOT NULL OR estado IN(en_proceso, ...). Antes
            -- las Fijas (eos_plan) que el usuario arrastró pero no inició
            -- NO aparecían en las cards · pérdida de visibilidad real.
            -- Ahora muestra TODO lo programado del horizonte (excepto
            -- canceladas) · estado='programado' es válido.
            ORDER BY pp.fecha_programada, pp.id
        """, (fechas_horizonte[0].isoformat(), fechas_horizonte[-1].isoformat())).fetchall()
        productos_db_por_fecha = set()  # (fecha, producto_upper)
        for r in rows_db:
            (pid, prod_n, kg, lotes, estado, area_id, f_iso,
             area_cod, area_nom,
             op_disp, op_elab, op_env, op_acon, inicio, fin) = r
            productos_db_por_fecha.add((f_iso, (prod_n or '').upper()))
            estado = (estado or 'programado').strip()
            # Acción según estado
            if estado == 'completado':
                accion, accion_label = None, '✅ Completada'
            elif estado in ('en_proceso', 'iniciado'):
                accion, accion_label = 'terminar', '✓ Terminar'
            elif area_id and op_disp:
                accion, accion_label = 'iniciar', '▶ Iniciar'
            else:
                accion, accion_label = 'asignar_ia', '🤖 IA asignar'
            producciones_dia.append({
                'id': pid, 'producto': prod_n, 'kg': kg, 'lotes': lotes or 1,
                'estado': estado, 'fecha': f_iso,
                'area': {'codigo': area_cod or '', 'nombre': area_nom or ''},
                'operarios': {
                    'dispensacion': (op_disp or '').strip(),
                    'elaboracion': (op_elab or '').strip(),
                    'envasado': (op_env or '').strip(),
                    'acondicionamiento': (op_acon or '').strip(),
                },
                'accion': accion, 'accion_label': accion_label,
                'desde_calendar': False,
                'inicio_real_at': inicio, 'fin_real_at': fin,
            })

        # 2) Eventos Calendar del horizonte que no están en DB → preview con IA
        # Sebastián 1-may-2026: 'mala lectura · solo hay 2 producciones, no 7'
        # FIX: filtrar SOLO eventos que parecen producciones (Fabricación,
        # Producción, código SKU). Los Comité/Cita/Etapa/Ritual NO son
        # producciones aunque el matcher pueda encontrarles partial match.
        try:
            from blueprints.auto_plan import (
                _calendar_events_cached, _match_producto_evento, _parsear_kg_evento
            )
            import re as _re
            cal_events = _calendar_events_cached(force_refresh=False) or []
            skus_aliases = {}
            for sku_n, alias_csv in conn.execute("""
                SELECT producto_nombre, COALESCE(alias_calendar,'')
                FROM sku_planeacion_config
                WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
            """).fetchall():
                skus_aliases[sku_n] = alias_csv

            # Patrones que indican PRODUCCIÓN real (no reuniones/comités)
            PROD_PATTERNS = _re.compile(
                r'(fabric|produc|elabor|envas|maquila)',
                _re.IGNORECASE
            )
            # Patrones que indican NO-producción (Comité, Cita, etc.)
            NO_PROD_PATTERNS = _re.compile(
                r'^\s*(comité|comite|cita|reunión|reunion|etapa|ritual|kickoff|'
                r'inventario|rrhh|capacitación|capacitacion|ducha|fumigación|'
                r'fumigacion|comité|alerta|booster\s+tensor|programado\s*:|'
                r'generar.*actas|decisión|decision|lanzami)',
                _re.IGNORECASE
            )

            # Deduplicación por (fecha, producto_final) Calendar
            productos_calendar_por_fecha = set()

            for ev in cal_events:
                try:
                    f_ev = ev.get('fecha', '')[:10]
                    if f_ev not in fechas_horizonte_iso: continue
                except Exception:
                    continue
                titulo_orig = (ev.get('titulo') or '').strip()
                # FILTRO 1: rechazar eventos que claramente no son producciones
                if NO_PROD_PATTERNS.search(titulo_orig):
                    continue
                # FILTRO 2: aceptar solo si tiene match SKU O patrón producción
                tiene_match_sku = False
                producto_match = None
                best = 0
                for prod_n, alias_csv in skus_aliases.items():
                    try:
                        s = _match_producto_evento(prod_n, alias_csv, ev.get('titulo'), ev.get('descripcion',''))
                        if s >= 50 and s > best:
                            best = s; producto_match = prod_n
                    except Exception:
                        continue
                if producto_match:
                    producto_final = producto_match
                    tiene_match_sku = True
                else:
                    # Sin match SKU · solo aceptar si título indica fabricación
                    if not PROD_PATTERNS.search(titulo_orig):
                        continue  # ej. "Comité Mensual" no es producción
                    titulo = titulo_orig
                    titulo = _re.sub(r'\s*[-–]\s*Fab(rica|ric)?[a-z]*\s+\d.*$', '', titulo, flags=_re.IGNORECASE)
                    titulo = _re.sub(r'\s*\(.*?\)\s*$', '', titulo)
                    titulo = _re.sub(r'\s*\d+\s*kg.*$', '', titulo, flags=_re.IGNORECASE)
                    producto_final = titulo.strip().upper() or 'EVENTO SIN TÍTULO'
                # Skip si ya está en DB para esa fecha (no duplicar)
                if (f_ev, producto_final.upper()) in productos_db_por_fecha:
                    continue
                # DEDUP entre eventos Calendar del mismo día con mismo producto
                key_cal = (f_ev, producto_final.upper())
                if key_cal in productos_calendar_por_fecha:
                    continue
                productos_calendar_por_fecha.add(key_cal)
                kg = _parsear_kg_evento(ev.get('titulo'), ev.get('descripcion','')) or 0
                # CM-FIX #3 · 20-may-2026 OLA Funcional · _ops_para()
                # INVENTABA operarios por hash determinístico · NO eran
                # persistidos · usuario veía "Mayerlin · Camilo" en card
                # que NUNCA se guardó. Ahora mostramos vacíos · cuando el
                # operario abre la card y aprieta "Iniciar (Calendar)" se
                # crea la fila DB con sus operarios reales (auto-asignación
                # honesta, no inventada).
                producciones_dia.append({
                    'id': None, 'producto': producto_final,
                    'kg': kg, 'lotes': 1,
                    'estado': 'planeado', 'fecha': f_ev,
                    'area': {'codigo': '', 'nombre': ''},
                    'operarios': {
                        'dispensacion': '', 'elaboracion': '',
                        'envasado': '', 'acondicionamiento': '',
                    },
                    'accion': 'iniciar_calendar',
                    'accion_label': '▶ Iniciar (Calendar)',
                    'desde_calendar': True,
                    'operarios_no_persistidos': True,  # señal a UI
                    'tiene_match_sku': tiene_match_sku,
                    'titulo_calendar': titulo_orig[:120],
                    'payload_iniciar': {
                        'producto': producto_final, 'fecha': f_ev,
                        'kg': kg, 'titulo': titulo_orig[:200],
                    },
                })
        except Exception as _e:
            logging.getLogger('programacion').warning(f'[centro-mando] preview Calendar falla: {_e}')

        # Ordenar por fecha + estado (en_proceso primero, completado al final)
        def _orden_key(p):
            est_orden = {'en_proceso': 0, 'iniciado': 0, 'programado': 1,
                         'planeado': 2, 'completado': 3}
            return (p.get('fecha', ''), est_orden.get(p.get('estado', ''), 4))
        producciones_dia.sort(key=_orden_key)

        # ── SCHEDULING SECUENCIAL (Sebastián 1-may-2026) ──
        # 'capacidad simultánea 4 · si hay más, recomendar secuencial'
        # 4 operarios pueden llevar 4 producciones paralelas (cada uno en
        # un rol distinto). Días con > 4 prods necesitan ondas secuenciales.
        CAPACIDAD_PARALELA = 4
        DURACION_DEFAULT_MIN = 180  # 3h por producción (estándar)
        HORA_INICIO = 7  # 7am
        # Agrupar por fecha
        prods_por_fecha = {}
        for p in producciones_dia:
            f = p.get('fecha', '')
            if f: prods_por_fecha.setdefault(f, []).append(p)
        # Calcular wave + horario sugerido por cada producción
        capacidad_warnings = []
        for f, prods in prods_por_fecha.items():
            num = len(prods)
            for idx, p in enumerate(prods):
                # Estimar duración: 3h base, +1h si >50kg, +2h si >100kg
                kg = p.get('kg') or 0
                dur_min = DURACION_DEFAULT_MIN
                if kg > 100: dur_min = 300
                elif kg > 50: dur_min = 240
                wave = (idx // CAPACIDAD_PARALELA) + 1  # 1, 2, 3...
                slot_min = (wave - 1) * DURACION_DEFAULT_MIN
                slot_h = HORA_INICIO + slot_min // 60
                slot_m = slot_min % 60
                fin_h = slot_h + dur_min // 60
                fin_m = (slot_min + dur_min) % 60
                p['wave'] = wave
                p['slot_inicio_sugerido'] = f'{slot_h:02d}:{slot_m:02d}'
                p['slot_fin_sugerido'] = f'{fin_h:02d}:{fin_m:02d}'
                p['duracion_estimada_min'] = dur_min
            if num > CAPACIDAD_PARALELA:
                capacidad_warnings.append({
                    'fecha': f,
                    'num_producciones': num,
                    'capacidad': CAPACIDAD_PARALELA,
                    'ondas_secuenciales': (num + CAPACIDAD_PARALELA - 1) // CAPACIDAD_PARALELA,
                    'extras_secuenciales': num - CAPACIDAD_PARALELA,
                })

        # ── DIAGNÓSTICO: filas DB sin match en Calendar (orphans/stale)
        # Sebastián 1-may-2026: 'siento que deja todo como lunes, no sabe
        # discriminar junta todo' → detectar duplicados de auto-sync antiguo
        # PERF FIX 24-may PM · auditoría agente · este pass iteraba
        # cal_events × skus_aliases (~24K ops) duplicando el trabajo
        # ya hecho en el primer pass (línea 3848). Reusar _productos_cal
        # si está disponible · solo reconstruir si skip_cleanup.
        try:
            productos_calendar_por_fecha = _productos_cal  # reuse
        except NameError:
            productos_calendar_por_fecha = set()
            try:
                from blueprints.auto_plan import (
                    _calendar_events_cached as _cec2,
                    _match_producto_evento as _mpe2,
                )
                _ac2 = conn.cursor()
                skus_aliases2 = {}
                for sku_n, alias_csv in _ac2.execute("""
                    SELECT producto_nombre, COALESCE(alias_calendar,'')
                    FROM sku_planeacion_config
                    WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
                """).fetchall():
                    skus_aliases2[sku_n] = alias_csv
                cal_events_check = _cec2(force_refresh=False) or []
                for ev in cal_events_check:
                    f_ev = (ev.get('fecha') or '')[:10]
                    if f_ev not in fechas_horizonte_iso:
                        continue
                    producto_match_d = None
                    best_d = 0
                    for prod_n, alias_csv in skus_aliases2.items():
                        try:
                            s = _mpe2(prod_n, alias_csv, ev.get('titulo'),
                                       ev.get('descripcion', ''))
                            if s >= 50 and s > best_d:
                                best_d = s
                                producto_match_d = prod_n
                        except Exception:
                            continue
                    if producto_match_d:
                        productos_calendar_por_fecha.add(
                            (f_ev, producto_match_d.upper()))
            except Exception:
                pass
        db_sin_calendar = []
        for p in producciones_dia:
            if p.get('desde_calendar'): continue  # estos vienen de Calendar
            if p.get('estado') in ('en_proceso','iniciado','completado'): continue  # ya iniciada
            key = (p.get('fecha',''), (p.get('producto','') or '').upper())
            if productos_calendar_por_fecha and key not in productos_calendar_por_fecha:
                db_sin_calendar.append({
                    'id': p.get('id'),
                    'producto': p.get('producto'),
                    'fecha': p.get('fecha'),
                    'kg': p.get('kg'),
                    'area_codigo': (p.get('area') or {}).get('codigo',''),
                })
    except Exception as _e:
        logging.getLogger('programacion').warning(f'[centro-mando] producciones_dia falla: {_e}')

    # Auto-limpieza ya se ejecutó al INICIO del endpoint (variables
    # auto_canceladas_inicial / auto_cancel_detalle_inicial) · garantiza
    # que producciones_dia construido arriba ya no incluye huérfanas
    auto_canceladas = auto_canceladas_inicial
    auto_cancel_detalle = auto_cancel_detalle_inicial

    return jsonify({
        'areas': areas,
        'kpis': {
            'producciones_activas_ahora': kpi_row[0] or 0,
            'terminadas_hoy': kpi_row[1] or 0,
            'cycle_time_promedio_min': int(kpi_row[2]) if kpi_row[2] else None,
            'salas_libres': salas_libres,
            'salas_sucias': salas_sucias,
            'salas_ocupadas': salas_ocupadas,
            'producciones_dia_total': len(producciones_dia),
            'producciones_dia_pendientes': sum(1 for p in producciones_dia
                                                  if p['estado'] in ('planeado','programado')),
            'auto_canceladas': auto_canceladas,
            # CM-FIX #6 · 20-may-2026: explicitar fecha de los KPIs.
            # 'activas_ahora' y 'terminadas_hoy' SIEMPRE son LIVE (no
            # respetan fecha_sel) · evita confusión cuando el usuario
            # navega días futuros desde el selector.
            'kpis_fecha': hoy,
            'kpis_son_live': True,
        },
        'producciones_dia': producciones_dia,
        'producciones_diag': {
            'db_sin_calendar': db_sin_calendar,
            'db_sin_calendar_count': len(db_sin_calendar),
            'auto_canceladas_esta_carga': auto_canceladas,
            'auto_cancel_detalle': auto_cancel_detalle[:5],
            'auto_clean_diag': auto_clean_diag,
            'capacidad_warnings': capacidad_warnings,
        },
        'eventos_recientes': eventos,
        'fecha': hoy,
    })


@bp.route('/api/planta/limpiar-db-sin-calendar', methods=['POST'])
def limpiar_db_sin_calendar():
    """Cancela filas produccion_programada que no tienen match en Calendar
    (en horizonte 7 días). Sebastián 1-may-2026: 'siento que deja todo como
    lunes · no sabe discriminar · junta todo'.

    Auto-sync antiguo dejó duplicados con fechas erróneas. Este endpoint
    los marca como 'cancelado' (NO delete · permite revertir si necesario).
    Solo afecta filas no iniciadas (estado in ('programado','planeado','')).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    body = request.json or {}
    fecha_param = (body.get('fecha') or '').strip()[:10]
    try:
        fecha_sel = datetime.strptime(fecha_param, '%Y-%m-%d').date() if fecha_param else date.today()
    except Exception:
        fecha_sel = date.today()
    fechas_horizonte = [fecha_sel + timedelta(days=i) for i in range(7)]
    fechas_iso = set(f.isoformat() for f in fechas_horizonte)

    conn = get_db(); c = conn.cursor()
    # 1) Construir set Calendar (fecha, producto_upper)
    productos_cal = set()
    try:
        from blueprints.auto_plan import _calendar_events_cached, _match_producto_evento
        cal_events = _calendar_events_cached(force_refresh=True) or []
        skus_aliases = {}
        for sku_n, alias_csv in c.execute("""
            SELECT producto_nombre, COALESCE(alias_calendar,'')
            FROM sku_planeacion_config
            WHERE activo=1 AND COALESCE(estado,'activo') NOT IN ('descontinuado','pausado')
        """).fetchall():
            skus_aliases[sku_n] = alias_csv
        for ev in cal_events:
            try:
                f_ev = ev.get('fecha', '')[:10]
                if f_ev not in fechas_iso: continue
            except Exception:
                continue
            producto_match = None
            best = 0
            for prod_n, alias_csv in skus_aliases.items():
                try:
                    s = _match_producto_evento(prod_n, alias_csv, ev.get('titulo'), ev.get('descripcion',''))
                    if s >= 50 and s > best:
                        best = s; producto_match = prod_n
                except Exception:
                    continue
            if producto_match:
                productos_cal.add((f_ev, producto_match.upper()))
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Calendar fetch falla: {e}'}), 500

    # 2) DB rows en horizonte sin match Calendar (no iniciadas)
    # Sebastián 19-may-2026: respetar lo FIJO · esta limpieza solo aplica
    # a sugerencias (canónico/calendar/manual), nunca a lo que el usuario fijó.
    rows = c.execute("""
        SELECT id, producto, date(fecha_programada) as f, COALESCE(estado,'')
        FROM produccion_programada
        WHERE date(fecha_programada) BETWEEN ? AND ?
          AND COALESCE(estado, 'programado') IN ('', 'programado', 'planeado')
          AND COALESCE(origen,'') NOT IN ('eos_plan','eos_b2b','eos_retroactivo')
    """, (fechas_horizonte[0].isoformat(), fechas_horizonte[-1].isoformat())).fetchall()

    a_cancelar = []
    for pid, prod, f, _est in rows:
        key = (f, (prod or '').upper())
        if productos_cal and key not in productos_cal:
            a_cancelar.append((pid, prod, f))

    if not a_cancelar:
        return jsonify({
            'ok': True,
            'mensaje': '✅ Sin filas DB huérfanas · todo está sincronizado con Calendar',
            'cancelados': 0,
            'productos_calendar_total': len(productos_cal),
        })

    # 3) Marcar como cancelado (no delete · easy revert)
    canceladas = 0
    for pid, prod, f in a_cancelar:
        try:
            c.execute("""
                UPDATE produccion_programada
                  SET estado='cancelado',
                      observaciones = COALESCE(observaciones,'') ||
                        ' [cancelado por limpieza · sin match Calendar · usuario=' || ? || ']'
                WHERE id=?
            """, (user, pid))
            canceladas += 1
            # AUDITORÍA-FIX 23-may-2026 · C16
            _sls = _liberar_sols_pre_produccion(c, pid, motivo='Producción cancelada por limpieza Calendar-first')
            try:
                audit_log(c, usuario=user, accion='AUTO_CANCELAR_PRODUCCION',
                          tabla='produccion_programada', registro_id=pid,
                          despues={'razon': 'limpieza · sin match Calendar',
                                   'producto': prod, 'fecha': f,
                                   'sols_liberadas': _sls})
            except Exception:
                pass
        except Exception:
            continue
    conn.commit()
    return jsonify({
        'ok': True,
        'mensaje': f'🗑 {canceladas} filas DB marcadas cancelado · sin match en Calendar',
        'cancelados': canceladas,
        'detalle': [{'producto': p, 'fecha': f} for _, p, f in a_cancelar[:20]],
        'productos_calendar_total': len(productos_cal),
    })


@bp.route('/api/programacion/mees-disponibles', methods=['GET'])
def mees_disponibles_lista():
    """Lista MEEs activos para alimentar dropdown del modal del lote.

    Sebastián 25-may-2026 PM · "que me salga desplegable solo en eso" ·
    el input texto libre causaba errores ('150ML' no existe). Ahora el
    frontend lee de aquí y muestra opciones reales del maestro.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    items = []
    try:
        for r in conn.execute(
            """SELECT codigo, COALESCE(descripcion,''),
                      COALESCE(categoria,''), COALESCE(stock_actual,0)
               FROM maestro_mee
               WHERE COALESCE(estado,'Activo') = 'Activo'
               ORDER BY categoria, descripcion, codigo""").fetchall():
            items.append({
                'codigo': r[0],
                'descripcion': r[1] or '',
                'categoria': r[2] or '',
                'stock_actual': float(r[3] or 0),
                'label': f'{r[0]} · {(r[1] or "")[:80]}' + (f' ({r[2]})' if r[2] else ''),
            })
    except Exception as e:
        return jsonify({'error': str(e)[:200], 'items': []}), 500
    return jsonify({'items': items, 'total': len(items)})


@bp.route('/api/programacion/lote/<int:lote_id>/envase-override', methods=['PATCH'])
def patch_envase_override_lote(lote_id):
    """Edita envase_codigo_override de un lote · admin elige envase
    distinto al default del producto (sku_mee_config).

    Sebastián 25-may-2026 PM · "en calendario faltaría poder agregarle
    el envase para empezar a calcular esas necesidades".

    Body: {envase_codigo_override: str}  ("" para limpiar y volver al default)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    envase = (d.get('envase_codigo_override') or '').strip().upper()
    conn = get_db(); cur = conn.cursor()
    lote = cur.execute(
        "SELECT id, producto, COALESCE(envase_codigo_override,''), "
        "COALESCE(estado,''), inicio_real_at, fin_real_at "
        "FROM produccion_programada WHERE id = ?", (lote_id,)).fetchone()
    if not lote:
        return jsonify({'error': 'Lote no existe'}), 404
    if (lote[3] or '').lower() in ('cancelado','completado'):
        return jsonify({'error': f'Lote {lote[3]} · no se puede editar'}), 400
    if lote[4] or lote[5]:
        return jsonify({'error': 'Lote ya inició/terminó · no se puede editar'}), 400
    if envase:
        # Validar que el envase exista en maestro_mee
        try:
            ok = cur.execute(
                "SELECT 1 FROM maestro_mee WHERE UPPER(TRIM(codigo)) = ?",
                (envase,)).fetchone()
            if not ok:
                return jsonify({'error': f'Envase {envase} no existe en maestro_mee',
                                 'codigo': 'ENVASE_NO_EXISTE'}), 400
        except Exception:
            pass
    antes_env = lote[2] or ''
    try:
        cur.execute(
            "UPDATE produccion_programada SET envase_codigo_override = ? WHERE id = ?",
            (envase, lote_id))
    except Exception as e:
        emsg = str(e).lower()
        if 'no such column' in emsg or 'does not exist' in emsg:
            return jsonify({'error': 'Migración 184 no aplicada · contactar admin'}), 500
        return jsonify({'error': str(e)[:200]}), 500
    try:
        from audit_helpers import audit_log as _al
        _al(cur, usuario=user, accion='PATCH_ENVASE_OVERRIDE_LOTE',
            tabla='produccion_programada', registro_id=lote_id,
            antes={'envase_codigo_override': antes_env},
            despues={'envase_codigo_override': envase})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'lote_id': lote_id,
                     'envase_codigo_override': envase,
                     'mensaje': ('✓ Envase ' + envase + ' asignado al lote · MEE recalculará')
                                 if envase else
                                 '✓ Envase override limpiado · vuelve al default del producto'})


@bp.route('/api/programacion/lote/<int:lote_id>/envase-aplicar-default', methods=['POST'])
def aplicar_envase_como_default_producto(lote_id):
    """Opción B · cambia el envase DEFAULT del producto (sku_mee_config)
    al envase override del lote · TODOS los lotes futuros del producto
    usarán este envase por default a menos que se setee otro override.

    Sebastián 25-may-2026 PM · "necesito a b y c". Toma el
    envase_codigo_override del lote · busca los SKUs del producto vía
    sku_producto_map · actualiza el item componente_tipo='envase' de
    cada SKU en sku_mee_config para apuntar al nuevo envase.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); cur = conn.cursor()
    lote = cur.execute(
        "SELECT id, producto, COALESCE(envase_codigo_override,''), "
        "COALESCE(estado,''), inicio_real_at, fin_real_at "
        "FROM produccion_programada WHERE id = ?", (lote_id,)).fetchone()
    if not lote:
        return jsonify({'error': 'Lote no existe'}), 404
    envase = (lote[2] or '').strip().upper()
    if not envase:
        return jsonify({'error': 'Lote no tiene envase override · setealo primero con el botón Guardar arriba'}), 400
    producto = (lote[1] or '').strip()
    if not producto:
        return jsonify({'error': 'Lote sin producto'}), 400
    # Validar envase en maestro_mee
    try:
        ok = cur.execute(
            "SELECT 1 FROM maestro_mee WHERE UPPER(TRIM(codigo)) = ?",
            (envase,)).fetchone()
        if not ok:
            return jsonify({'error': f'Envase {envase} no existe en maestro_mee'}), 400
    except Exception:
        pass
    # Buscar SKUs del producto vía sku_producto_map
    skus = []
    try:
        for r in cur.execute(
            """SELECT UPPER(TRIM(sku)) FROM sku_producto_map
               WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
                 AND COALESCE(activo,1) = 1""", (producto,)).fetchall():
            if r[0]:
                skus.append(r[0])
    except Exception:
        pass
    if not skus:
        return jsonify({'error': f'Producto {producto} no tiene SKUs mapeados en sku_producto_map · imposible cambiar default'}), 404
    # Para cada SKU, actualizar el item de tipo envase (o insertar si no existe)
    afectados = []
    for sku in skus:
        # Ver si ya hay un componente_tipo='envase' para este sku
        env_actual_row = cur.execute(
            """SELECT id, mee_codigo FROM sku_mee_config
               WHERE UPPER(TRIM(sku_codigo)) = ?
                 AND componente_tipo = 'envase'
                 AND COALESCE(aplica,1) = 1 LIMIT 1""",
            (sku,)).fetchone()
        if env_actual_row:
            cur.execute(
                "UPDATE sku_mee_config SET mee_codigo = ? WHERE id = ?",
                (envase, env_actual_row[0]))
            afectados.append({'sku': sku, 'envase_antes': env_actual_row[1],
                              'envase_nuevo': envase, 'accion': 'UPDATE'})
        else:
            try:
                cur.execute(
                    """INSERT INTO sku_mee_config
                         (sku_codigo, mee_codigo, componente_tipo, cantidad_por_unidad, aplica)
                       VALUES (?, ?, 'envase', 1, 1)""",
                    (sku, envase))
                afectados.append({'sku': sku, 'envase_antes': '',
                                  'envase_nuevo': envase, 'accion': 'INSERT'})
            except Exception:
                pass
    # audit_log · uno por SKU para trazabilidad completa (audit fix
    # del audit 25-may-2026 PM · antes solo guardaba detalle[:5])
    try:
        from audit_helpers import audit_log as _al
        # Audit resumen primero
        _al(cur, usuario=user, accion='ENVASE_DEFAULT_GLOBAL_PRODUCTO',
            tabla='sku_mee_config', registro_id=lote_id,
            despues={'producto': producto, 'envase_nuevo': envase,
                      'skus_afectados': len(afectados),
                      'detalle_completo': afectados})
    except Exception:
        pass
    conn.commit()
    return jsonify({
        'ok': True, 'producto': producto, 'envase_default_nuevo': envase,
        'skus_afectados': len(afectados), 'detalle': afectados,
        'mensaje': f'✓ Envase default de {producto} cambiado a {envase} en {len(afectados)} SKU(s) · futuros lotes lo usan automático',
    })


@bp.route('/api/programacion/lote/<int:lote_id>/envase-propagar-futuros', methods=['POST'])
def propagar_envase_a_futuros(lote_id):
    """Opción C · aplica el envase override del lote a TODOS los lotes
    FUTUROS del mismo producto que aún no iniciaron · sin tocar el
    default global (sku_mee_config queda igual).

    Útil cuando: el cambio aplica solo a los próximos N lotes pendientes
    pero el default permanente del producto sigue siendo otro.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); cur = conn.cursor()
    lote = cur.execute(
        "SELECT id, producto, fecha_programada, COALESCE(envase_codigo_override,'') "
        "FROM produccion_programada WHERE id = ?", (lote_id,)).fetchone()
    if not lote:
        return jsonify({'error': 'Lote no existe'}), 404
    envase = (lote[3] or '').strip().upper()
    if not envase:
        return jsonify({'error': 'Lote no tiene envase override · setealo primero'}), 400
    producto = (lote[1] or '').strip()
    fecha = (lote[2] or '')[:10]
    if not producto or not fecha:
        return jsonify({'error': 'Lote sin producto o fecha'}), 400
    # Validar envase
    try:
        ok = cur.execute(
            "SELECT 1 FROM maestro_mee WHERE UPPER(TRIM(codigo)) = ?",
            (envase,)).fetchone()
        if not ok:
            return jsonify({'error': f'Envase {envase} no existe'}), 400
    except Exception:
        pass
    # Update lotes futuros · mismo producto · fecha > este · no iniciados
    # · estado no terminal · excluye el propio lote
    # Sebastián 25-may-2026 PM · audit bug fix · NO sobreescribir lotes
    # B2B dedicados (origen='eos_b2b') · esos tienen envase pactado con el
    # cliente en pedidos_b2b_lote.envase_codigo · propagar Animus DTC encima
    # rompería la promesa al cliente. eos_retroactivo también excluido por
    # seguridad (datos del pasado · no se tocan).
    try:
        cur.execute(
            """UPDATE produccion_programada
               SET envase_codigo_override = ?
               WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
                 AND fecha_programada > ?
                 AND id != ?
                 AND inicio_real_at IS NULL
                 AND fin_real_at IS NULL
                 AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
                 AND COALESCE(origen,'') NOT IN ('eos_b2b','eos_retroactivo')""",
            (envase, producto, fecha, lote_id))
        afectados = cur.rowcount or 0
        # Contar cuántos B2B dedicados quedaron sin propagar (para info al admin)
        b2b_no_propagados = cur.execute(
            """SELECT COUNT(*) FROM produccion_programada
               WHERE UPPER(TRIM(producto)) = UPPER(TRIM(?))
                 AND fecha_programada > ?
                 AND id != ?
                 AND inicio_real_at IS NULL AND fin_real_at IS NULL
                 AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
                 AND COALESCE(origen,'') IN ('eos_b2b','eos_retroactivo')""",
            (producto, fecha, lote_id)).fetchone()
        b2b_skip = int(b2b_no_propagados[0] or 0) if b2b_no_propagados else 0
    except Exception as e:
        return jsonify({'error': f'UPDATE fallo: {e}'}), 500
    try:
        from audit_helpers import audit_log as _al
        _al(cur, usuario=user, accion='ENVASE_PROPAGAR_FUTUROS',
            tabla='produccion_programada', registro_id=lote_id,
            despues={'producto': producto, 'envase': envase,
                      'desde_fecha': fecha, 'lotes_actualizados': afectados,
                      'lotes_b2b_excluidos': b2b_skip})
    except Exception: pass
    conn.commit()
    msg = f'✓ {afectados} lote(s) futuros de {producto} actualizados con envase {envase}'
    if b2b_skip > 0:
        msg += f' · ⚠ {b2b_skip} lote(s) B2B/retroactivo NO modificados (mantienen envase pactado)'
    return jsonify({
        'ok': True, 'producto': producto, 'envase': envase,
        'desde_fecha': fecha, 'lotes_actualizados': afectados,
        'lotes_b2b_excluidos': b2b_skip,
        'mensaje': msg,
    })


@bp.route('/api/programacion/programar/<int:evento_id>/composicion-mee', methods=['GET'])
def prog_composicion_mee(evento_id):
    """Devuelve la composición de envases que una producción específica va a
    consumir, derivada de producto_presentaciones + ratio Shopify 90d.

    Sebastián 27-may-2026 PM · "el calendario debe colocar la realidad del
    envase, que lo tenga cada producción". Auto-derive desde Shopify · no
    manual. Visible en /calendario y /planta cuando se inicia producción.

    Response:
    {
      ok, producto, cantidad_kg,
      variantes: [
        {presentacion_codigo, etiqueta, volumen_ml, envase_codigo,
         envase_descripcion, ratio_pct, unidades_estimadas}, ...
      ],
      fuente_ratio: 'shopify_90d' | 'uniforme' | 'unica',
      sin_variantes: bool   // True si producto no tiene producto_presentaciones
    }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    conn = db_connect(); c = conn.cursor()
    _res = _composicion_envases_lote(c, evento_id)
    if _res is None:
        return jsonify({'error': f'Producción {evento_id} no encontrada'}), 404
    # Desglose por cliente × envase (DTC = composición − B2B) · Sebastián 1-jun-2026
    try:
        _res['plan_por_cliente'] = _plan_envasado_por_cliente(
            c, evento_id, _res.get('variantes') or [])
    except Exception:
        _res['plan_por_cliente'] = []
    return jsonify(_res)


def _ventas_sku_180d(c):
    """Ventas por SKU (Shopify) últimos 180d · memoizado por request en flask.g.
    FIX 1-jun-2026 (audit escalabilidad): antes se re-escaneaba animus_shopify_orders
    + se parseaba JSON por CADA producción (N+1 O(N·M)). Ahora se calcula 1 vez."""
    _g = None
    try:
        from flask import g as _g
        _cached = getattr(_g, '_ventas_sku_180d_cache', None)
        if _cached is not None:
            return _cached
    except Exception:
        _g = None
    from datetime import datetime as _dtV, timedelta as _tdV
    import json as _jsonV
    cutoff = (_dtV.utcnow() - _tdV(hours=5) - _tdV(days=180)).date().isoformat()
    ventas = {}
    try:
        for (it,) in c.execute(
            """SELECT sku_items FROM animus_shopify_orders
               WHERE COALESCE(creado_en,'') >= ? AND sku_items IS NOT NULL
                 AND sku_items != ''""", (cutoff,)).fetchall():
            try:
                items = _jsonV.loads(it) if it else []
                if not isinstance(items, list):
                    continue
                for li in items:
                    sk = (li.get('sku') or '').strip()
                    qty = int(li.get('qty') or 0)
                    if sk and qty > 0:
                        ventas[sk] = ventas.get(sk, 0) + qty
            except Exception:
                continue
    except sqlite3.OperationalError:
        pass
    if _g is not None:
        try:
            _g._ventas_sku_180d_cache = ventas
        except Exception:
            pass
    return ventas


def _composicion_envases_lote(c, evento_id):
    """Composición de envases (variantes) de un lote · helper reusable por
    composicion-mee y por la preparación de envases. Devuelve dict o None
    (si el lote no existe). Misma lógica · sin duplicar."""
    # fija_override_json (mig 205) · override de cantidad fija POR LOTE · fallback
    # si la columna no existe (instancia sin migrar).
    row = None
    for _ovr in ("COALESCE(fija_override_json,'')", "''"):
        try:
            row = c.execute(
                "SELECT id, producto, cantidad_kg, lotes, " + _ovr +
                " FROM produccion_programada WHERE id=?", (evento_id,)).fetchone()
            break
        except Exception:
            row = None
    if not row:
        return None
    pid, producto, cant_kg, lotes = row[0], row[1], row[2], row[3]
    # Override por lote · {presentacion_codigo: uds}
    fija_override = {}
    try:
        import json as _jovr
        _raw = row[4] if len(row) > 4 else ''
        if _raw:
            _parsed = _jovr.loads(_raw)
            if isinstance(_parsed, dict):
                fija_override = {str(k).strip().upper(): float(v)
                                 for k, v in _parsed.items() if v is not None}
    except Exception:
        fija_override = {}
    producto_norm = (producto or '').strip().upper()
    cant_kg_f = float(cant_kg or 0)

    # Cargar presentaciones activas del producto. cantidad_fija_uds (mig 204)
    # puede no existir en instancias viejas → fallback graceful a 0.
    presentaciones = []
    _pp_base = (
        "presentacion_codigo, COALESCE(etiqueta,''), COALESCE(volumen_ml,0), "
        "COALESCE(envase_codigo,''), COALESCE(sku_shopify,''), "
        "COALESCE(ventas_mes_referencia,0), {fija} "
        "FROM producto_presentaciones "
        "WHERE LOWER(TRIM(producto_nombre))=LOWER(TRIM(?)) "
        "AND COALESCE(activo,1)=1 AND COALESCE(volumen_ml,0) > 0 "
        "ORDER BY volumen_ml DESC")
    _rows_pp = None
    for _fija_expr in ("COALESCE(cantidad_fija_uds,0)", "0"):
        try:
            _rows_pp = c.execute(
                "SELECT " + _pp_base.format(fija=_fija_expr), (producto,)).fetchall()
            break
        except Exception:
            _rows_pp = None
            continue
    for r in (_rows_pp or []):
        presentaciones.append({
            'codigo': r[0], 'etiqueta': r[1],
            'volumen_ml': float(r[2]), 'envase_codigo': r[3],
            'sku_shopify': r[4],
            'ventas_mes_referencia': float(r[5] or 0),
            'cantidad_fija_uds': float(r[6] or 0),
        })

    if not presentaciones:
        # Fallback · solo 1 volumen sin presentaciones definidas
        vol = 0
        try:
            vr = c.execute(
                "SELECT COALESCE(volumen_ml,0) FROM volumen_unitario_producto WHERE LOWER(TRIM(producto_nombre))=LOWER(TRIM(?)) AND COALESCE(activo,1)=1",
                (producto,)
            ).fetchone()
            vol = float((vr or [0])[0] or 0)
        except Exception:
            pass
        units = int(round((cant_kg_f * 1000.0) / vol)) if vol > 0 else 0
        return {
            'ok': True, 'producto': producto, 'cantidad_kg': cant_kg_f,
            'variantes': [{
                'presentacion_codigo': '-', 'etiqueta': f'{int(vol)} ml' if vol > 0 else '(sin volumen)',
                'volumen_ml': vol, 'envase_codigo': '',
                'envase_descripcion': '',
                'ratio_pct': 100.0, 'unidades_estimadas': units,
            }] if vol > 0 else [],
            'fuente_ratio': 'unica', 'sin_variantes': True,
        }

    # Ratio Shopify · ventana ampliada a 180d (6 meses) Sebastián 27-may-2026 PM
    # · "si lanzas la busqueda de ventas en 6 meses sabras que se vendio no
    # puede se 50 50". Capturar ciclo de venta más realista para SKUs B2B
    # cosmético con cadencia mensual.
    ventas_sku = _ventas_sku_180d(c)  # memoizado por request (FIX N+1 · 1-jun-2026)

    # Derivar ratio por presentación · prioridad: 1) override manual,
    # 2) Shopify histórico 180d, 3) uniforme
    suma_override = sum(float(p.get('ventas_mes_referencia') or 0) for p in presentaciones)
    if suma_override > 0:
        for p in presentaciones:
            p['_ventas_90d'] = float(p.get('ventas_mes_referencia') or 0)
            p['_ratio'] = (p['_ventas_90d'] / suma_override)
        fuente = 'manual_override'
    else:
        total = 0
        for p in presentaciones:
            sk = (p['sku_shopify'] or '').strip()
            v = ventas_sku.get(sk, 0) if sk else 0
            p['_ventas_90d'] = v
            total += v
        if total > 0:
            for p in presentaciones:
                p['_ratio'] = (p['_ventas_90d'] / total)
            fuente = 'shopify_180d'
        else:
            ratio_uni = 1.0 / len(presentaciones)
            for p in presentaciones:
                p['_ratio'] = ratio_uni
            fuente = 'uniforme'

    # Calcular unidades estimadas + descripción envase
    mee_descs = {}
    try:
        for r in c.execute("SELECT codigo, COALESCE(descripcion,'') FROM maestro_mee").fetchall():
            mee_descs[r[0]] = r[1]
    except Exception:
        pass

    # Sebastián 30-may-2026 · presentaciones con CANTIDAD FIJA (mig 204):
    # se reservan PRIMERO sus uds (y su kg); el bulk RESTANTE se reparte por
    # ratio entre las NO-fijas. Caso TRX: 10ml fijo 1200 uds, resto al 30ml.
    kg_fijo_total = 0.0
    for p in presentaciones:
        _cod = (p['codigo'] or '').strip().upper()
        _ovr = fija_override.get(_cod)            # override POR LOTE (mig 205)
        _default = float(p.get('cantidad_fija_uds') or 0)  # default del producto
        fija = _ovr if _ovr is not None else _default      # override gana
        p['_fija_efectiva'] = fija
        p['_fija_default'] = _default
        p['_fija_es_override'] = _ovr is not None
        p['_es_fija'] = fija > 0 and p['volumen_ml'] > 0
        p['_kg_fijo'] = (fija * p['volumen_ml']) / 1000.0 if p['_es_fija'] else 0.0
        if p['_es_fija']:
            kg_fijo_total += p['_kg_fijo']
    # Si lo fijo excede el bulk, escalar hacia abajo para no pasarse.
    fija_scale = (cant_kg_f / kg_fijo_total) if (kg_fijo_total > cant_kg_f and kg_fijo_total > 0) else 1.0
    kg_restante = max(0.0, cant_kg_f - min(kg_fijo_total, cant_kg_f))
    suma_ratio_no_fija = sum(p['_ratio'] for p in presentaciones if not p['_es_fija'])
    n_no_fija = sum(1 for p in presentaciones if not p['_es_fija'])
    hay_fija = any(p['_es_fija'] for p in presentaciones)

    variantes_out = []
    for p in presentaciones:
        if p['_es_fija']:
            kg_p = p['_kg_fijo'] * fija_scale
            un_p = int(round((kg_p * 1000.0) / p['volumen_ml'])) if p['volumen_ml'] > 0 else 0
        else:
            if suma_ratio_no_fija > 0:
                r_rel = p['_ratio'] / suma_ratio_no_fija
            else:
                r_rel = (1.0 / n_no_fija) if n_no_fija else 0.0
            kg_p = kg_restante * r_rel
            un_p = int(round((kg_p * 1000.0) / p['volumen_ml'])) if p['volumen_ml'] > 0 else 0
        variantes_out.append({
            'presentacion_codigo': p['codigo'],
            'etiqueta': p['etiqueta'],
            'volumen_ml': p['volumen_ml'],
            'envase_codigo': p['envase_codigo'],
            'envase_descripcion': mee_descs.get(p['envase_codigo'], p['envase_codigo']),
            'sku_shopify': p['sku_shopify'],
            'ratio_pct': round((kg_p / cant_kg_f * 100), 1) if cant_kg_f > 0 else 0,
            'ventas_90d_uds': p['_ventas_90d'],
            'cantidad_fija_uds': float(p.get('_fija_efectiva') or 0),
            'cantidad_fija_default': float(p.get('_fija_default') or 0),
            'fija_es_override': bool(p.get('_fija_es_override')),
            'es_fija': p['_es_fija'],
            'unidades_estimadas': un_p,
        })

    return {
        'ok': True, 'producto': producto, 'cantidad_kg': cant_kg_f,
        'variantes': variantes_out,
        'fuente_ratio': ('fija+' + fuente) if hay_fija else fuente,
        'tiene_fija': hay_fija,
        'tiene_override': bool(fija_override),
        'sin_variantes': False,
    }


def _plan_envasado_por_cliente(c, evento_id, variantes):
    """Desglose del envasado POR CLIENTE × ENVASE (Sebastián 1-jun-2026: "no deja
    claro cuánto envasar de 10/30 para Animus y cuántos de 30 para Kelly").

    Modelo (decisión Sebastián): el B2B SALE de la mezcla → DTC = composición − B2B
    por volumen. Cada aporte B2B (pedidos_b2b_lote) toma sus unidades de su envase
    (match por ml); lo que queda de cada variante es de Animus DTC.

    Devuelve: [{cliente, es_dtc, envases:[{etiqueta, ml, uds, es_fija}]}, ...]
    Animus DTC primero, luego cada cliente B2B."""
    aportes = []
    try:
        for r in c.execute(
            "SELECT COALESCE(cliente_nombre,''), COALESCE(ml_unidad,0), "
            "COALESCE(unidades_aporte,0), COALESCE(envase_codigo,'') "
            "FROM pedidos_b2b_lote WHERE lote_produccion_id=? "
            "ORDER BY kg_aporte DESC", (evento_id,)).fetchall():
            aportes.append({'cliente': (r[0] or 'B2B'), 'ml': round(float(r[1] or 0), 1),
                            'uds': int(r[2] or 0), 'envase': r[3] or ''})
    except Exception:
        aportes = []  # mig 171/172 no aplicada

    # uds B2B por ml (para restar del DTC)
    b2b_uds_por_ml = {}
    for a in aportes:
        b2b_uds_por_ml[a['ml']] = b2b_uds_por_ml.get(a['ml'], 0) + a['uds']

    dtc_envases = []
    for v in (variantes or []):
        ml = round(float(v.get('volumen_ml') or 0), 1)
        uds_total = int(v.get('unidades_estimadas') or 0)
        uds_dtc = max(0, uds_total - int(b2b_uds_por_ml.get(ml, 0)))
        if uds_total > 0:
            dtc_envases.append({
                'etiqueta': v.get('etiqueta') or (f'{int(ml)}ml' if ml else '—'),
                'ml': ml, 'uds': uds_dtc, 'es_fija': bool(v.get('es_fija')),
            })

    out = []
    if any(e['uds'] > 0 for e in dtc_envases):
        out.append({'cliente': 'Animus DTC', 'es_dtc': True, 'envases': dtc_envases})
    for a in aportes:
        out.append({
            'cliente': a['cliente'], 'es_dtc': False,
            'envases': [{
                'etiqueta': (f"{int(a['ml'])}ml" if a['ml'] else (a['envase'] or '—')),
                'ml': a['ml'], 'uds': a['uds'], 'es_fija': False,
            }],
        })
    return out


@bp.route('/api/compras/preparacion-envases', methods=['GET'])
def compras_preparacion_envases():
    """Jalona los envases de las producciones próximas (no iniciadas) para que
    Compras escoja cuáles preparar (serigrafía/tampografía). Reusa el cálculo de
    composición de envases. fecha_lista = fecha_produccion − anticipo (30d def).
    Read-only · no muta. Sebastián 31-may-2026 (Pieza 1)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    from datetime import date as _dpe, timedelta as _tdpe, datetime as _dtpe
    try:
        dias = max(7, min(365, int(request.args.get('dias', 90))))
    except Exception:
        dias = 90
    try:
        anticipo = max(0, min(180, int(request.args.get('anticipo', 30))))
    except Exception:
        anticipo = 30
    conn = get_db(); c = conn.cursor()
    hoy = (_dtpe.utcnow() - _tdpe(hours=5)).date()
    hasta = (hoy + _tdpe(days=dias)).isoformat()
    rows = c.execute(
        """SELECT id, producto, fecha_programada FROM produccion_programada
           WHERE COALESCE(estado,'') NOT IN ('cancelado','completado')
             AND inicio_real_at IS NULL
             AND date(fecha_programada) >= date('now','-5 hours')
             AND date(fecha_programada) <= date(?)
           ORDER BY fecha_programada""", (hasta,)).fetchall()
    # OS ya existentes por envase (para avisar y no duplicar)
    os_por_envase = {}
    try:
        for r in c.execute(
            "SELECT envase_codigo_mee, COUNT(*) FROM ordenes_servicio "
            "WHERE COALESCE(estado,'') != 'Cancelada' GROUP BY envase_codigo_mee").fetchall():
            if r[0]:
                os_por_envase[r[0]] = int(r[1] or 0)
    except Exception:
        pass
    items = []
    for (pid, producto, fecha) in rows:
        comp = _composicion_envases_lote(c, pid)
        if not comp:
            continue
        f10 = (fecha or '')[:10]
        try:
            fl = (_dpe.fromisoformat(f10) - _tdpe(days=anticipo)).isoformat()
        except Exception:
            fl = None
        atrasado = bool(fl and fl < hoy.isoformat())
        for v in (comp.get('variantes') or []):
            env = (v.get('envase_codigo') or '').strip()
            if not env:
                continue
            items.append({
                'lote_id': pid, 'producto': producto,
                'fecha_produccion': f10, 'fecha_lista_sugerida': fl,
                'lista_atrasada': atrasado,
                'envase_codigo': env,
                'envase_descripcion': v.get('envase_descripcion') or env,
                'presentacion': v.get('etiqueta'),
                'volumen_ml': v.get('volumen_ml'),
                'uds': int(v.get('unidades_estimadas') or 0),
                'os_existentes': os_por_envase.get(env, 0),
            })
    items.sort(key=lambda x: (x.get('fecha_lista_sugerida') or '9999'))
    return jsonify({'ok': True, 'dias': dias, 'anticipo_dias': anticipo,
                    'items': items, 'n': len(items)})


@bp.route('/api/compras/minimos-envases-sugeridos', methods=['GET'])
def compras_minimos_envases_sugeridos():
    """Mínimo de stock SUGERIDO de cada envase (MEE) según el consumo REAL del
    plan de producción (no el estático 1000). minimo = consumo_diario × cobertura.
    Read-only · compara vs el stock_minimo actual. Sebastián 31-may-2026 (Pieza 3)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    from datetime import datetime as _dm, timedelta as _tdm
    import math as _math
    try:
        dias = max(30, min(365, int(request.args.get('dias', 90))))
    except Exception:
        dias = 90
    try:
        cobertura = max(7, min(180, int(request.args.get('cobertura_dias', 45))))
    except Exception:
        cobertura = 45
    conn = get_db(); c = conn.cursor()
    hoy = (_dm.utcnow() - _tdm(hours=5)).date()
    hasta = (hoy + _tdm(days=dias)).isoformat()
    rows = c.execute(
        """SELECT id FROM produccion_programada
           WHERE COALESCE(estado,'') NOT IN ('cancelado','completado')
             AND inicio_real_at IS NULL
             AND date(fecha_programada) >= date('now','-5 hours')
             AND date(fecha_programada) <= date(?)""", (hasta,)).fetchall()
    consumo = {}
    for (pid,) in rows:
        comp = _composicion_envases_lote(c, pid)
        if not comp:
            continue
        for v in (comp.get('variantes') or []):
            env = (v.get('envase_codigo') or '').strip()
            if not env:
                continue
            consumo[env] = consumo.get(env, 0) + int(v.get('unidades_estimadas') or 0)
    mee = {}
    try:
        for r in c.execute(
            "SELECT codigo, COALESCE(descripcion,''), COALESCE(stock_minimo,0), "
            "COALESCE(stock_actual,0) FROM maestro_mee").fetchall():
            mee[r[0]] = {'descripcion': r[1], 'stock_minimo': float(r[2] or 0),
                         'stock_actual': float(r[3] or 0)}
    except Exception:
        pass
    items = []
    for env, uds in consumo.items():
        diario = uds / float(dias)
        sugerido = int(_math.ceil(diario * cobertura))
        info = mee.get(env, {})
        actual = info.get('stock_minimo', 0)
        items.append({
            'envase_codigo': env, 'descripcion': info.get('descripcion', ''),
            'consumo_horizonte': uds, 'consumo_diario': round(diario, 2),
            'minimo_actual': actual, 'minimo_sugerido': sugerido,
            'stock_actual': info.get('stock_actual', 0),
            'diff': sugerido - actual, 'en_maestro': env in mee,
        })
    items.sort(key=lambda x: -x['consumo_horizonte'])
    return jsonify({'ok': True, 'dias': dias, 'cobertura_dias': cobertura,
                    'items': items, 'n': len(items)})


@bp.route('/api/compras/minimos-envases-aplicar', methods=['POST'])
def compras_minimos_envases_aplicar():
    """Aplica stock_minimo a los MEE elegidos. Body {items:[{codigo, stock_minimo}]}."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    items = d.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'error': 'items requerido'}), 400
    conn = get_db(); c = conn.cursor()
    n = 0
    for it in items[:500]:
        cod = (it.get('codigo') or '').strip()
        if not cod:
            continue
        try:
            sm = max(0.0, float(it.get('stock_minimo')))
        except (TypeError, ValueError):
            continue
        cur = c.execute("UPDATE maestro_mee SET stock_minimo=? WHERE codigo=?", (sm, cod))
        if cur.rowcount:
            n += 1
    try:
        audit_log(c, usuario=user, accion='MINIMOS_MEE_APLICAR',
                  tabla='maestro_mee', registro_id='', despues={'n_actualizados': n})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'actualizados': n})


@bp.route('/api/programacion/lote/<int:lote_id>/fija-override', methods=['PATCH'])
def prog_fija_override(lote_id):
    """Override de cantidad FIJA de una presentación SOLO para este lote (mig 205).

    Sebastián 30-may-2026: el default vive en producto_presentaciones (modal admin);
    esto permite ajustar un lote puntual (ej. promo: 2000 minis en vez de 1200) sin
    tocar el default. Body: {presentacion_codigo, uds}. uds null/'' → quita el
    override (vuelve al default del producto). No muta producto_presentaciones.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    cod = (d.get('presentacion_codigo') or '').strip().upper()
    if not cod:
        return jsonify({'error': 'presentacion_codigo requerido'}), 400
    uds_raw = d.get('uds', None)
    import json as _jfo
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT COALESCE(fija_override_json,''), COALESCE(inicio_real_at,''), "
        "COALESCE(fin_real_at,'') FROM produccion_programada WHERE id=?",
        (lote_id,)).fetchone()
    if not row:
        return jsonify({'error': f'Lote {lote_id} no existe'}), 404
    if row[1] or row[2]:
        return jsonify({'error': 'Lote ya iniciado/terminado · no se puede cambiar'}), 400
    try:
        ovr = _jfo.loads(row[0]) if row[0] else {}
        if not isinstance(ovr, dict):
            ovr = {}
    except Exception:
        ovr = {}
    if uds_raw is None or str(uds_raw).strip() == '':
        ovr.pop(cod, None)            # quitar override → vuelve al default
        accion = 'quitado'
    else:
        try:
            v = max(0.0, float(uds_raw))
        except (TypeError, ValueError):
            return jsonify({'error': 'uds inválido'}), 400
        ovr[cod] = v
        accion = f'fijado {v:g}'
    nuevo = _jfo.dumps(ovr) if ovr else None
    try:
        c.execute("UPDATE produccion_programada SET fija_override_json=? WHERE id=?",
                  (nuevo, lote_id))
    except Exception as e:
        return jsonify({'error': f'no se pudo guardar: {e}'}), 500
    try:
        audit_log(c, usuario=user, accion='FIJA_OVERRIDE_LOTE',
                  tabla='produccion_programada', registro_id=lote_id,
                  despues={'presentacion': cod, 'override': ovr})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'lote_id': lote_id, 'presentacion_codigo': cod,
                    'override': ovr,
                    'mensaje': f'Override {accion} para este lote'})


@bp.route('/api/programacion/lote/<int:lote_id>/plan-envasado/<int:pbl_id>', methods=['PATCH'])
def patch_plan_envasado_b2b(lote_id, pbl_id):
    """Edita plan_envasado_uds y plan_envasado_notas de un cliente B2B
    específico en un lote.

    Sebastián 25-may-2026 PM · "deberías colocar que yo mismo lo escriba,
    y tenga algo como observaciones". Por cada cliente del lote, admin
    puede sobreescribir las unidades a envasar (en vez del cálculo auto
    kg*1000/ml) y agregar observaciones libres para el operario.

    Body: {plan_envasado_uds?: int, plan_envasado_notas?: str}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json(silent=True) or {}
    conn = get_db(); cur = conn.cursor()
    # Validar que el link existe y corresponde al lote
    row = cur.execute(
        """SELECT id, pedido_b2b_id, lote_produccion_id, cliente_nombre,
                  COALESCE(plan_envasado_uds, 0), COALESCE(plan_envasado_notas, '')
           FROM pedidos_b2b_lote WHERE id = ? AND lote_produccion_id = ?""",
        (pbl_id, lote_id)).fetchone()
    if not row:
        return jsonify({'error': 'Link pedido-lote no existe o no corresponde'}), 404
    sets, params = [], []
    cambios = {}
    if 'plan_envasado_uds' in d:
        try:
            uds = int(d.get('plan_envasado_uds') or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'plan_envasado_uds debe ser entero'}), 400
        if uds < 0 or uds > 10_000_000:
            return jsonify({'error': 'plan_envasado_uds fuera de rango'}), 400
        sets.append('plan_envasado_uds = ?')
        params.append(uds)
        cambios['plan_envasado_uds'] = uds
    if 'plan_envasado_notas' in d:
        notas = (d.get('plan_envasado_notas') or '').strip()[:500]
        sets.append('plan_envasado_notas = ?')
        params.append(notas)
        cambios['plan_envasado_notas'] = notas
    if not sets:
        return jsonify({'error': 'nada que actualizar'}), 400
    params.append(pbl_id)
    cur.execute(
        f"UPDATE pedidos_b2b_lote SET {', '.join(sets)} WHERE id = ?", params)
    try:
        from audit_helpers import audit_log as _al
        _al(cur, usuario=user, accion='PATCH_PLAN_ENVASADO_B2B',
            tabla='pedidos_b2b_lote', registro_id=pbl_id,
            antes={'plan_envasado_uds': row[4], 'plan_envasado_notas': row[5]},
            despues=cambios)
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'pbl_id': pbl_id, 'cambios': cambios})


@bp.route('/api/programacion/produccion-programada/listado', methods=['GET'])
def listado_produccion_programada():
    """Lista todas las producciones programadas activas con su origen,
    para diagnostico de producciones fantasma. Sebastian lo necesita
    cuando ve algo en el horizonte que no recuerda haber programado.

    Sebastián 25-may-2026 PM · cada lote ahora incluye `desglose_b2b`:
    array de {cliente_nombre, kg_aporte, unidades_aporte, modo} leído
    de pedidos_b2b_lote. El frontend calcula `kg_dtc = kg_total - sum(b2b)`.
    Si el resultado es negativo o cero, significa que el lote fue
    "inflado a ojo" en el pasado y la diferencia no está atribuida.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    # Sebastián 9-jun-2026: ?historico=1 → muestra TODO el histórico (incl. completadas,
    # sin ventana de fecha · "debe permanecer todo, es el histórico"). Las completadas son
    # de fechas pasadas → solo llenan celdas pasadas, no ensucian el plan adelante. Default
    # (sin flag) = vista de plan (últimos 7 días + futuro, sin completadas · diag admin).
    _historico = (request.args.get('historico') or '').strip().lower() in ('1', 'true', 'si')
    if _historico:
        _where_clause = ("WHERE LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado') "
                         "ORDER BY pp.fecha_programada ASC, pp.id ASC")
    else:
        _where_clause = ("WHERE LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado') "
                         "AND pp.fecha_programada >= date('now', '-5 hours', '-7 day') "
                         "ORDER BY pp.fecha_programada ASC, pp.id ASC")
    # Sebastián 25-may-2026 PM · agregar envase_codigo_override (mig 184)
    # Fallback al SELECT sin la columna si la mig no aplicó aún
    try:
        rows = conn.execute("""
            SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
                   pp.estado, COALESCE(pp.origen,'manual') as origen,
                   pp.observaciones,
                   COALESCE(pp.cantidad_kg,
                            pp.lotes * COALESCE(fh.lote_size_kg, 0),
                            0) as kg,
                   pp.area_id, ap.nombre as area_nombre,
                   pp.operario_dispensacion_id, od.nombre || ' ' || COALESCE(od.apellido,'') as op_disp,
                   pp.operario_elaboracion_id,  oe.nombre || ' ' || COALESCE(oe.apellido,'') as op_elab,
                   pp.operario_envasado_id,     oen.nombre || ' ' || COALESCE(oen.apellido,'') as op_env,
                   pp.operario_acondicionamiento_id, oa.nombre || ' ' || COALESCE(oa.apellido,'') as op_acon,
                   COALESCE(pp.envase_codigo_override, '') as envase_override
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            LEFT JOIN areas_planta ap ON ap.id = pp.area_id
            LEFT JOIN operarios_planta od  ON od.id  = pp.operario_dispensacion_id
            LEFT JOIN operarios_planta oe  ON oe.id  = pp.operario_elaboracion_id
            LEFT JOIN operarios_planta oen ON oen.id = pp.operario_envasado_id
            LEFT JOIN operarios_planta oa  ON oa.id  = pp.operario_acondicionamiento_id
        """ + " " + _where_clause).fetchall()
        _has_env_ovr = True
    except Exception:
        rows = conn.execute("""
            SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
                   pp.estado, COALESCE(pp.origen,'manual') as origen,
                   pp.observaciones,
                   COALESCE(pp.cantidad_kg,
                            pp.lotes * COALESCE(fh.lote_size_kg, 0),
                            0) as kg,
                   pp.area_id, ap.nombre as area_nombre,
                   pp.operario_dispensacion_id, od.nombre || ' ' || COALESCE(od.apellido,'') as op_disp,
                   pp.operario_elaboracion_id,  oe.nombre || ' ' || COALESCE(oe.apellido,'') as op_elab,
                   pp.operario_envasado_id,     oen.nombre || ' ' || COALESCE(oen.apellido,'') as op_env,
                   pp.operario_acondicionamiento_id, oa.nombre || ' ' || COALESCE(oa.apellido,'') as op_acon
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            LEFT JOIN areas_planta ap ON ap.id = pp.area_id
            LEFT JOIN operarios_planta od  ON od.id  = pp.operario_dispensacion_id
            LEFT JOIN operarios_planta oe  ON oe.id  = pp.operario_elaboracion_id
            LEFT JOIN operarios_planta oen ON oen.id = pp.operario_envasado_id
            LEFT JOIN operarios_planta oa  ON oa.id  = pp.operario_acondicionamiento_id
        """ + " " + _where_clause).fetchall()
        _has_env_ovr = False
    lote_ids = [r[0] for r in rows]
    desglose_por_lote = {}
    if lote_ids:
        try:
            placeholders = ','.join('?' * len(lote_ids))
            # Sebastián 25-may-2026 PM · agregar plan_envasado_uds + notas
            # (mig 183) · si no aplicada, fallback al SELECT sin esos campos
            try:
                b2b_rows = conn.execute(
                    f"""SELECT pbl.lote_produccion_id,
                              COALESCE(NULLIF(TRIM(pbl.cliente_nombre),''),
                                       NULLIF(TRIM(pb.cliente_nombre),''),
                                       pb.cliente_id, 'B2B') as cliente,
                              COALESCE(pbl.kg_aporte, 0) as kg_aporte,
                              COALESCE(pbl.unidades_aporte, 0) as uds_aporte,
                              COALESCE(pbl.modo, 'sumado_a_lote_canonico') as modo,
                              pbl.pedido_b2b_id,
                              COALESCE(pbl.envase_codigo, '') as envase,
                              COALESCE(pbl.ml_unidad, 0) as ml,
                              COALESCE(pbl.plan_envasado_uds, 0) as plan_uds,
                              COALESCE(pbl.plan_envasado_notas, '') as plan_notas,
                              pbl.id as pbl_id
                        FROM pedidos_b2b_lote pbl
                        LEFT JOIN pedidos_b2b pb ON pb.id = pbl.pedido_b2b_id
                        WHERE pbl.lote_produccion_id IN ({placeholders})
                        ORDER BY pbl.lote_produccion_id, kg_aporte DESC""",
                    lote_ids).fetchall()
                _has_plan = True
            except Exception:
                b2b_rows = conn.execute(
                    f"""SELECT pbl.lote_produccion_id,
                              COALESCE(NULLIF(TRIM(pbl.cliente_nombre),''),
                                       NULLIF(TRIM(pb.cliente_nombre),''),
                                       pb.cliente_id, 'B2B') as cliente,
                              COALESCE(pbl.kg_aporte, 0) as kg_aporte,
                              COALESCE(pbl.unidades_aporte, 0) as uds_aporte,
                              COALESCE(pbl.modo, 'sumado_a_lote_canonico') as modo,
                              pbl.pedido_b2b_id,
                              COALESCE(pbl.envase_codigo, '') as envase,
                              COALESCE(pbl.ml_unidad, 0) as ml,
                              pbl.id as pbl_id
                        FROM pedidos_b2b_lote pbl
                        LEFT JOIN pedidos_b2b pb ON pb.id = pbl.pedido_b2b_id
                        WHERE pbl.lote_produccion_id IN ({placeholders})
                        ORDER BY pbl.lote_produccion_id, kg_aporte DESC""",
                    lote_ids).fetchall()
                _has_plan = False
            for br in b2b_rows:
                kg_aporte = float(br[2] or 0)
                ml_u = float(br[7] or 0)
                # Unidades calculadas si no hay plan editado: kg*1000/ml
                uds_calc = 0
                if ml_u > 0:
                    uds_calc = int(round(kg_aporte * 1000 / ml_u))
                plan_uds = int(br[8] or 0) if _has_plan else 0
                plan_notas = (br[9] if _has_plan else '') or ''
                pbl_id = br[10] if _has_plan else br[8]
                desglose_por_lote.setdefault(br[0], []).append({
                    'cliente': br[1] or 'B2B',
                    'kg': kg_aporte,
                    'unidades': int(br[3] or 0),
                    'modo': br[4] or 'sumado_a_lote_canonico',
                    'pedido_id': br[5],
                    'envase': br[6] or '',
                    'ml': ml_u,
                    'unidades_calculadas': uds_calc,
                    'plan_envasado_uds': plan_uds,
                    'plan_envasado_notas': plan_notas,
                    'pbl_id': pbl_id,
                })
        except Exception as _e_b2b:
            # Tabla pedidos_b2b_lote no existe en bootstrap (mig 169 viejo)
            # · log + seguir sin desglose · no romper /admin/plan-calendario.
            log = logging.getLogger('inventario.programacion')
            log.warning('desglose B2B fallo: %s', _e_b2b)
    out = []
    for r in rows:
        kg_total = float(r[7] or 0)
        desglose = desglose_por_lote.get(r[0], [])
        kg_b2b_sum = sum(d['kg'] for d in desglose)
        kg_dtc = round(kg_total - kg_b2b_sum, 2)
        # "Lote inflado": tiene B2B y el DTC queda > 0 (caso normal). Si
        # kg_dtc < 0 → más B2B atribuido que el lote total · data
        # inconsistente · frontend muestra warning. Si kg_b2b_sum == 0
        # → todo DTC, sin desglose explícito (caso default).
        envase_override = (r[18] if _has_env_ovr and len(r) > 18 else '') or ''
        out.append({
            'id': r[0], 'producto': r[1], 'fecha_programada': r[2], 'lotes': r[3],
            'estado': r[4], 'origen': r[5], 'observaciones': r[6], 'kg': kg_total,
            'area_id': r[8], 'area_nombre': r[9],
            'operario_dispensacion_id': r[10], 'operario_dispensacion': (r[11] or '').strip() or None,
            'operario_elaboracion_id':  r[12], 'operario_elaboracion':  (r[13] or '').strip() or None,
            'operario_envasado_id':     r[14], 'operario_envasado':     (r[15] or '').strip() or None,
            'operario_acondicionamiento_id': r[16], 'operario_acondicionamiento': (r[17] or '').strip() or None,
            'envase_codigo_override': envase_override,
            'desglose_b2b': desglose,
            'kg_b2b_total': round(kg_b2b_sum, 2),
            'kg_dtc': max(0.0, kg_dtc),
            'split_inconsistente': kg_dtc < -0.01,
        })
    return jsonify({'producciones': out, 'total': len(out)})


@bp.route('/api/programacion/debug-calendar', methods=['GET'])
def debug_calendar_eventos():
    """Diagnóstico de TODOS los eventos del Calendar y por qué cada uno
    aparece (o NO aparece) en /producciones-faltantes.

    Sebastián 8-may-2026: la app solo muestra ~11 productos pero Calendar
    tiene 20+ eventos. La razón es silenciosa: si un SKU del título no
    está en sku_producto_map o no tiene fórmula, el evento se ignora.

    Este endpoint expone los 4 motivos por los que un evento no aparece:
      · sin_sku_detectable · regex no encontró código tipo [A-Z]+
      · sku_no_mapeado · el SKU detectado no está en sku_producto_map
      · sin_formula · el producto mapeado no tiene formula_headers
      · evento_no_fab · título contiene 'envasado'/'qc'/etc · skip por diseño
      · ok · evento sí aparece en la app

    Querystring: ?dias=14 (mismo horizonte que producciones-faltantes)

    Returns:
      {
        eventos: [{titulo, fecha, status, skus_detectados,
                   sku_match, producto_mapeado, tiene_formula, razon}],
        total_eventos: int, total_aparecen: int, total_ignorados: int,
        skus_no_mapeados: [list of SKUs unique],
        productos_sin_formula: [list of productos]
      }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    try:
        from blueprints.compras import ADMIN_USERS as _ADMIN
    except Exception:
        _ADMIN = ('sebastian', 'alejandro')
    if user not in _ADMIN:
        return jsonify({'error': 'Solo admin'}), 403

    import re as _re
    try:
        dias = max(7, min(int(request.args.get('dias', 14)), 365))
    except (ValueError, TypeError):
        dias = 14

    try:
        cal = _fetch_calendar_events(days_ahead=dias)
    except Exception as e:
        return jsonify({'error': f'Calendar fetch fallo: {e}'}), 500

    if cal.get('error'):
        return jsonify({
            'error': f'Calendar API: {cal["error"]}',
            'source': cal.get('source'),
            'hint': 'Configura GCAL_ICAL_URL en Render',
        }), 500

    events = cal.get('events') or []
    if not events:
        return jsonify({
            'eventos': [], 'total_eventos': 0,
            'note': 'Calendar respondió 0 eventos · ¿el calendar correcto?',
        })

    # Cargar mapeos relevantes
    conn = get_db()
    sku_to_prod = {}
    try:
        for r in conn.execute(
            "SELECT UPPER(sku), producto_nombre FROM sku_producto_map WHERE activo=1"
        ).fetchall():
            sku_to_prod[r[0]] = r[1]
    except Exception:
        pass

    productos_con_formula = set()
    try:
        for r in conn.execute(
            "SELECT UPPER(TRIM(producto_nombre)) FROM formula_headers"
        ).fetchall():
            productos_con_formula.add(r[0])
    except Exception:
        pass

    # Mismas constantes que el sync
    NOT_SKU = {
        'FAB','QC','CON','SIN','MICRO','ENVASADO','ACONDICIONAMIENTO',
        'FABRICACION','FABRICACIÓN','LANZAMIENTO','PRODUCCION','PRODUCCIÓN',
        'KG','MES','DIAS','DÍAS','ML','UDS','BATCH','FERNANDO',
        'MESA','LOTES','LOTE','MINI','MACRO','AND','THE','FOR',
        'FAB1','FYE2','FYE3','ENV1','ENV2','PROD1','PROD2','PROD3','PROD4'
    }
    NON_FAB_KW = NON_FAB_KW_GLOBAL  # Sebastián 12-may-2026: unificado

    def _skus(titulo):
        tokens = _re.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', (titulo or '').upper())
        return [t for t in tokens if t not in NOT_SKU]

    eventos_out = []
    skus_no_mapeados = set()
    productos_sin_formula = set()
    counters = {'ok': 0, 'sku_no_mapeado': 0, 'sin_formula': 0,
                'sin_sku': 0, 'no_fab': 0}

    for ev in events:
        titulo = ev.get('titulo', '') or ''
        fecha = ev.get('fecha', '')
        tlow = titulo.lower()

        # Status del evento
        if any(kw in tlow for kw in NON_FAB_KW):
            status = 'no_fab'
            razon = 'Evento de envasado/QC · ignorado por diseño'
            counters['no_fab'] += 1
            sku_match = None
            producto_mapeado = None
            tiene_formula = False
            skus_detectados = []
        else:
            skus_detectados = _skus(titulo)
            if not skus_detectados:
                status = 'sin_sku'
                razon = 'Título no tiene SKU detectable (regex [A-Z][A-Z0-9]+)'
                counters['sin_sku'] += 1
                sku_match = None
                producto_mapeado = None
                tiene_formula = False
            else:
                # Buscar primer SKU que matchee
                sku_match = None
                producto_mapeado = None
                tiene_formula = False
                for sku in skus_detectados:
                    if sku in sku_to_prod:
                        sku_match = sku
                        producto_mapeado = sku_to_prod[sku]
                        prod_norm = (producto_mapeado or '').strip().upper()
                        tiene_formula = prod_norm in productos_con_formula
                        break
                if not sku_match:
                    status = 'sku_no_mapeado'
                    razon = (f'SKU(s) {skus_detectados} no están en '
                             'sku_producto_map. Agregar mapping desde admin.')
                    counters['sku_no_mapeado'] += 1
                    for s in skus_detectados:
                        skus_no_mapeados.add(s)
                elif not tiene_formula:
                    status = 'sin_formula'
                    razon = (f'SKU={sku_match} mapea a producto "{producto_mapeado}" '
                             'pero NO tiene formula_headers · MPs no se calculan.')
                    counters['sin_formula'] += 1
                    productos_sin_formula.add(producto_mapeado)
                else:
                    status = 'ok'
                    razon = 'Evento aparece en /producciones-faltantes'
                    counters['ok'] += 1

        eventos_out.append({
            'titulo': titulo,
            'fecha': fecha,
            'skus_detectados': skus_detectados,
            'sku_match': sku_match,
            'producto_mapeado': producto_mapeado,
            'tiene_formula': tiene_formula,
            'status': status,
            'razon': razon,
        })

    return jsonify({
        'horizonte_dias': dias,
        'total_eventos': len(events),
        'eventos': eventos_out,
        'counters': counters,
        'skus_no_mapeados': sorted(skus_no_mapeados),
        'productos_sin_formula': sorted(productos_sin_formula),
        'hint': (
            'SKUs no mapeados: agregalos en admin / sku_producto_map. '
            'Productos sin fórmula: cargá la fórmula en /planta → fórmulas. '
            'Mientras tanto el evento queda invisible en la app.'
        ),
    })


@bp.route('/api/programacion/debug-producto/<path:producto>', methods=['GET'])
def debug_producto_estado(producto):
    """Diagnóstico de un producto: entries DB + eventos Calendar + razones
    de por qué el sync no las puede tocar.

    Sebastián 7-may-2026 (caso AZHC Lun 11): después de re-sync espejo,
    una entry sigue ahí · ¿por qué? Este endpoint responde:
      · Lista de entries en produccion_programada para el producto
        (id, fecha, lotes, kg, estado, origen, inicio_real_at,
         inventario_descontado_at, observaciones)
      · Eventos en Google Calendar que mencionan el SKU
      · Si el sync las borraría/cancelaría o no (guard explicado)

    Returns:
      { producto, entries_db: [...], eventos_calendar: [...],
        sku_to_prod_match: bool, formula_existe: bool }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    try:
        from blueprints.compras import ADMIN_USERS as _ADMIN
    except Exception:
        _ADMIN = ('sebastian', 'alejandro')
    if user not in _ADMIN:
        return jsonify({'error': 'Solo admin'}), 403

    producto_norm = (producto or '').strip().upper()
    conn = get_db(); c = conn.cursor()

    # Entries DB
    entries = []
    try:
        rows = c.execute("""
            SELECT id, producto, fecha_programada,
                   COALESCE(lotes, 1), COALESCE(cantidad_kg, 0),
                   COALESCE(estado, 'programado'),
                   COALESCE(origen, 'manual'),
                   COALESCE(inicio_real_at, ''),
                   COALESCE(inventario_descontado_at, ''),
                   COALESCE(observaciones, ''),
                   COALESCE(area_id, 0)
            FROM produccion_programada
            WHERE UPPER(TRIM(producto)) = ?
            ORDER BY fecha_programada DESC
        """, (producto_norm,)).fetchall()
        for r in rows:
            (id_, prod, fecha, lotes, kg, estado, origen,
             inicio, descontado, obs, area_id) = r
            # Razón guard
            guard_blocks = []
            if inicio:
                guard_blocks.append(f'inicio_real_at={inicio}')
            if descontado:
                guard_blocks.append(f'inventario_descontado_at={descontado}')
            if estado.lower() in ('cancelado', 'completado'):
                guard_blocks.append(f'estado={estado}')
            entries.append({
                'id': id_,
                'producto': prod,
                'fecha_programada': fecha,
                'lotes': lotes,
                'cantidad_kg': kg,
                'estado': estado,
                'origen': origen,
                'inicio_real_at': inicio or None,
                'inventario_descontado_at': descontado or None,
                'observaciones': obs[:200] if obs else '',
                'area_id': area_id,
                'protegida_del_sync': bool(guard_blocks),
                'razones_guard': guard_blocks,
            })
    except Exception as e:
        return jsonify({'error': f'DB: {e}'}), 500

    # SKU lookup · ¿Se mapea este producto a algún SKU?
    sku_match = []
    try:
        sku_rows = c.execute(
            "SELECT sku, activo FROM sku_producto_map WHERE UPPER(producto_nombre)=?",
            (producto_norm,)
        ).fetchall()
        for s, act in sku_rows:
            sku_match.append({'sku': s, 'activo': bool(act)})
    except Exception:
        pass

    # ¿Existe formula?
    formula_existe = False
    try:
        row = c.execute(
            "SELECT 1 FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=? LIMIT 1",
            (producto_norm,)
        ).fetchone()
        formula_existe = bool(row)
    except Exception:
        pass

    # Eventos Calendar (best-effort · puede fallar si no hay credenciales)
    eventos_cal = []
    try:
        cal = _fetch_calendar_events(days_ahead=120)
        import re as _re
        for ev in cal.get('events', []):
            titulo = ev.get('titulo', '') or ''
            tokens = _re.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', titulo.upper())
            # Si algún SKU del map matchea, o el producto está en el título
            matches = [s['sku'] for s in sku_match if s['sku'] in tokens]
            prod_in_title = producto_norm.replace(' ', '') in titulo.upper().replace(' ', '')
            if matches or prod_in_title:
                eventos_cal.append({
                    'titulo': titulo,
                    'fecha': ev.get('fecha'),
                    'skus_matcheados': matches,
                    'producto_en_titulo': prod_in_title,
                })
    except Exception as e:
        eventos_cal = [{'error': str(e)}]

    return jsonify({
        'producto': producto_norm,
        'entries_db': entries,
        'entries_count': len(entries),
        'eventos_calendar': eventos_cal,
        'eventos_calendar_count': len([e for e in eventos_cal
                                         if 'error' not in e]),
        'sku_match': sku_match,
        'formula_existe': formula_existe,
        'hint': (
            'Si ves entries con protegida_del_sync=true · esas no las borra '
            'el espejo · usá DELETE /api/programacion/produccion-programada/<id>/borrar'
            ' o el botón 🗑️ en la UI. Si eventos_calendar está vacío o no '
            'matchea SKUs · el evento Calendar tiene un título sin SKU '
            'reconocible o el SKU no está en sku_producto_map.'
        ),
    })


@bp.route('/api/programacion/produccion-programada/<int:evento_id>/borrar', methods=['DELETE'])
def borrar_produccion_programada(evento_id):
    """Borra una produccion programada (HARD DELETE, no solo cancelar).
    Solo admin. Usado para limpiar fantasmas que aparecen en el horizonte
    sin razon (ej. una entrada manual antigua que sobrevivio)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    # Importar lista de admins desde compras (misma fuente de verdad)
    try:
        from blueprints.compras import ADMIN_USERS as _ADMIN
    except Exception:
        _ADMIN = ('sebastian', 'alejandro')
    if user not in _ADMIN:
        return jsonify({'error': 'Solo admin puede borrar'}), 403
    conn = get_db(); c = conn.cursor()
    # P0 audit 26-may-2026 · zero-error · CLAUDE.md: "audit_log mandatory en
    # cualquier UPDATE/DELETE de produccion_programada · Una cancelación/borrado
    # de produccion_programada que no auditó es la que hizo desaparecer la
    # programación del 19-may sin dejar rastro". Y guard contra borrar lotes
    # en curso o que ya descontaron inventario (drift silencioso).
    row = c.execute(
        "SELECT producto, fecha_programada, COALESCE(origen,'manual'), "
        "       cantidad_kg, inicio_real_at, inventario_descontado_at, estado "
        "FROM produccion_programada WHERE id=?",
        (evento_id,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Produccion no encontrada'}), 404
    force = (request.args.get('force') or '').lower() in ('1','true','yes')
    _es_fijo = (row[2] or '') in ('eos_plan', 'eos_b2b', 'eos_retroactivo')
    if (row[4] or row[5]) and not force:
        return jsonify({
            'error': 'Produccion en curso o con inventario ya descontado · no borrable sin ?force=1',
            'inicio_real_at': row[4],
            'inventario_descontado_at': row[5],
            'estado': row[6],
            'hint': 'Si insistes, reenvía con ?force=1 (queda auditado · drift de inventario posible)',
        }), 409
    # Sebastián 27-may-2026 · audit · borrar un FIJO (lo que el usuario fijó:
    # eos_plan/eos_b2b/eos_retroactivo) requiere force explícito · evita perder
    # planificación deliberada por un click accidental.
    if _es_fijo and not force:
        return jsonify({
            'error': f'Producción FIJA (origen={row[2]}) · es planificación deliberada · no borrable sin ?force=1',
            'origen': row[2],
            'estado': row[6],
            'hint': 'Si realmente querés borrar lo que se fijó manualmente, reenvía con ?force=1 (queda auditado)',
        }), 409
    # Snapshot antes para audit
    snap = {
        'producto': row[0], 'fecha_programada': row[1], 'origen': row[2],
        'cantidad_kg': row[3], 'inicio_real_at': row[4],
        'inventario_descontado_at': row[5], 'estado': row[6],
    }
    # Tambien borrar items de checklist huerfanos asociados
    try:
        c.execute("DELETE FROM produccion_checklist WHERE produccion_id=?", (evento_id,))
    except sqlite3.OperationalError:
        pass
    c.execute("DELETE FROM produccion_programada WHERE id=?", (evento_id,))
    # Audit log obligatorio (CLAUDE.md INV-6 · trail incident 19-may)
    try:
        from audit_helpers import audit_log as _al
        _al(c, usuario=user, accion='HARD_DELETE_PRODUCCION_PROGRAMADA',
            tabla='produccion_programada', registro_id=evento_id,
            antes=snap,
            detalle=f"HARD DELETE id={evento_id} producto={row[0]} origen={row[2]} force={force}")
    except Exception as _ae:
        import logging as _lg
        _lg.getLogger('programacion').warning('audit borrar PP fallo: %s', _ae)
    conn.commit()
    return jsonify({
        'ok': True, 'id': evento_id,
        'producto': row[0], 'fecha': row[1], 'origen': row[2],
        'force': force,
        'mensaje': f'Produccion {row[0]} ({row[1]}) borrada definitivamente',
    })


def _resolver_material_bodega(c, formula_mid, formula_nombre):
    """FIX 1-jun-2026 (P0 · frena producción) · resuelve el material_id de FÓRMULA al
    material_id que usa BODEGA/movimientos. Bug: al producir 'N-acetil glucosamina' decía
    'Hay 0g' aunque bodega tenía 600g bajo otro código (los ~116 MPs con ID distinto entre
    fórmula y bodega). La validación de stock + FEFO buscaban movimientos por el id de
    fórmula EXACTO, sin bridge/nombre/alias. Mismo patrón M1 que el resto del audit.

    Tiers (SOLO se resuelve si el id de fórmula NO tiene movimientos → los MPs que ya
    funcionan quedan IDÉNTICOS · cero riesgo de regresión):
      1) el propio formula_mid si ya tiene movimientos → se devuelve igual.
      2) mp_formula_bridge (formula_material_id → bodega_material_id).
      3) match por nombre normalizado/alias contra maestro_mps con movimientos.
    Devuelve el id resuelto, o el formula_mid original si no se encontró nada mejor."""
    fmid = str(formula_mid or '').strip()

    def _tiene_mov(mid):
        if not mid:
            return False
        try:
            return c.execute("SELECT 1 FROM movimientos WHERE material_id=? LIMIT 1", (mid,)).fetchone() is not None
        except Exception:
            return False

    def _stock_neto(mid):
        """Stock neto producible de un código (excluye estados no-producibles)."""
        if not mid:
            return 0.0
        try:
            _ph = ','.join(['?'] * len(_ESTADOS_LOTE_NO_PRODUCIBLES))
            r = c.execute(
                f"""SELECT COALESCE(SUM(CASE
                      WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                      WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END),0)
                    FROM movimientos WHERE material_id=?
                      AND UPPER(COALESCE(estado_lote,'')) NOT IN ({_ph})""",
                (mid,) + _ESTADOS_LOTE_NO_PRODUCIBLES).fetchone()
            return float(r[0] or 0)
        except Exception:
            return 0.0

    # 1) si el código de fórmula YA tiene stock neto producible → es el bueno.
    # Audit 4-jun · antes cortaba por _tiene_mov (cualquier movimiento histórico)
    # → un código con neto 0 (típico tras unify, o canónico consumido) devolvía 0g
    # y NUNCA llegaba al match que halla el código duplicado CON stock.
    if fmid and _stock_neto(fmid) > 0:
        return fmid
    # 2) bridge explícito (si el destino tiene stock)
    try:
        r = c.execute(
            "SELECT bodega_material_id FROM mp_formula_bridge "
            "WHERE TRIM(formula_material_id)=? AND COALESCE(activo,1)=1 LIMIT 1", (fmid,)).fetchone()
        if r and r[0] and _stock_neto(str(r[0]).strip()) > 0:
            return str(r[0]).strip()
    except Exception:
        pass
    # 2b) por INCI del código de fórmula · Audit 4-jun · robusto (no depende del
    # texto del nombre): si el código de fórmula tiene un INCI, buscar OTROS códigos
    # activos con el MISMO INCI y elegir el de más stock neto. Resuelve los
    # duplicados de un material (PANTHENOL en MP00110/MP00236) sin unificar.
    try:
        if fmid:
            _ir = c.execute("SELECT COALESCE(nombre_inci,'') FROM maestro_mps WHERE codigo_mp=?", (fmid,)).fetchone()
            _inci_f = _norm_mp_name(_ir[0]) if (_ir and _ir[0]) else ''
            if _inci_f:
                _inci_cands = []
                for r in c.execute("SELECT codigo_mp, COALESCE(nombre_inci,'') FROM maestro_mps WHERE COALESCE(activo,1)=1").fetchall():
                    cod = str(r[0] or '').strip()
                    if cod and _norm_mp_name(r[1]) == _inci_f and _stock_neto(cod) > 0:
                        _inci_cands.append(cod)
                if _inci_cands:
                    return sorted(_inci_cands, key=lambda x: (-_stock_neto(x), x))[0]
    except Exception:
        pass
    # 3) por nombre (exacto/normalizado/alias) contra maestro_mps.
    # Audit 4-jun · cuando el nombre matchea VARIOS códigos (mismo material en
    # códigos duplicados, p.ej. PANTHENOL en MP00110/MP00236), elegir el de MÁS
    # stock neto → producción jala el inventario aunque el código no esté unificado.
    # Es seguro porque nombres distintos = materiales distintos NO se matchean
    # (Polysorbate 80 ≠ 20, tetrapéptido ≠ glucosamina): solo compite el MISMO
    # material consigo mismo. Determinista (max stock, desempate por código).
    nom = (formula_nombre or '').strip()
    if nom:
        nn = _norm_mp_name(nom)
        _alias_nn = _MP_NAME_ALIAS.get(nn)
        try:
            _cands = set()
            for r in c.execute(
                "SELECT codigo_mp, COALESCE(nombre_inci,''), COALESCE(nombre_comercial,'') "
                "FROM maestro_mps WHERE COALESCE(activo,1)=1").fetchall():
                cod = str(r[0] or '').strip()
                if not cod:
                    continue
                for cand in (r[1], r[2]):
                    if not cand:
                        continue
                    cn = _norm_mp_name(cand)
                    if cn == nn or cn == _alias_nn or _MP_NAME_ALIAS.get(cn) == nn:
                        _cands.add(cod)
                        break
            if _cands:
                # preferir el de mayor stock neto; si ninguno tiene stock, el que
                # tenga movimientos; desempate determinista por código.
                _best = sorted(_cands, key=lambda x: (-_stock_neto(x),
                               0 if _tiene_mov(x) else 1, x))[0]
                if len(_cands) > 1:
                    logging.getLogger('programacion').warning(
                        "resolver MP nombre '%s' (fmid=%s) varios códigos %s → elige %s "
                        "(más stock) · unificá para limpiar", nom, fmid, sorted(_cands), _best)
                return _best
        except Exception:
            pass
    # 4) RESCATE FINAL · stock atrapado en código INACTIVO · Audit 4-jun
    # Caso PANTENOL: el código de fórmula (MP00236) y su duplicado (MP00110, 1118g)
    # están AMBOS inactivos tras una unificación a medias → tiers 2b/3 (solo activos)
    # devuelven 0g y producción aborta "no hay" aunque el inventario exista físico.
    # Aquí, como último recurso, buscamos por INCI o por nombre SIN filtrar activo y
    # elegimos el código con MÁS stock neto. Es inventario real y usable; el código
    # esté activo o no, el material existe en bodega. Log para que se limpie luego.
    try:
        _inci_f = ''
        if fmid:
            _ir = c.execute("SELECT COALESCE(nombre_inci,'') FROM maestro_mps WHERE codigo_mp=?", (fmid,)).fetchone()
            _inci_f = _norm_mp_name(_ir[0]) if (_ir and _ir[0]) else ''
        nn = _norm_mp_name(nom) if nom else ''
        _alias_nn = _MP_NAME_ALIAS.get(nn) if nn else None
        _resc = set()
        for r in c.execute(
            "SELECT codigo_mp, COALESCE(nombre_inci,''), COALESCE(nombre_comercial,'') "
            "FROM maestro_mps").fetchall():  # SIN filtro activo
            cod = str(r[0] or '').strip()
            if not cod or cod == (fmid or ''):
                continue
            _ok = False
            if _inci_f and _norm_mp_name(r[1]) == _inci_f:
                _ok = True
            elif nn:
                for cand in (r[1], r[2]):
                    if not cand:
                        continue
                    cn = _norm_mp_name(cand)
                    if cn == nn or cn == _alias_nn or _MP_NAME_ALIAS.get(cn) == nn:
                        _ok = True
                        break
            if _ok and _stock_neto(cod) > 0:
                _resc.add(cod)
        if _resc:
            _best = sorted(_resc, key=lambda x: (-_stock_neto(x), x))[0]
            logging.getLogger('programacion').warning(
                "resolver MP RESCATE inactivo · fórmula '%s' (fmid=%s) → %s tiene "
                "%.1fg atrapados (código posiblemente inactivo) · unificá para limpiar",
                nom or _inci_f, fmid, _best, _stock_neto(_best))
            return _best
    except Exception:
        pass
    return fmid  # fallback · comportamiento previo


def _calcular_mp_consumo_produccion(c, evento_id):
    """Calcula MPs a consumir por una producción programada.

    Lee formula_items y aplica:
      - Preferimos cantidad_g_por_lote * lotes (formula con cantidades fijas)
      - Fallback: porcentaje * cantidad_kg * 1000 (formula con %)

    Returns: (mps_a_consumir, meta) donde:
      mps_a_consumir = [{codigo_mp, nombre, cantidad_g}, ...]
      meta = {producto, fecha, lotes, cantidad_kg_total}
    Returns ({}, None) si la producción no existe.
    """
    row = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
               COALESCE(pp.cantidad_kg, 0)         as cantidad_kg_explicita,
               COALESCE(fh.lote_size_kg, 0)        as lote_kg_formula
        FROM produccion_programada pp
        LEFT JOIN formula_headers fh
               ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
        WHERE pp.id=?
    """, (evento_id,)).fetchone()
    if not row:
        return [], None
    pid, producto, fecha, lotes, cant_kg_exp, lote_kg = row
    lotes = int(lotes or 1)
    cant_kg_total = float(cant_kg_exp or 0) or (lotes * float(lote_kg or 0))

    rows = c.execute("""
        SELECT material_id, material_nombre,
               COALESCE(porcentaje, 0)                as pct,
               COALESCE(cantidad_g_por_lote, 0)       as g_por_lote
        FROM formula_items
        WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
    """, (producto,)).fetchall()
    # FIX 1-jun-2026 audit MP/fórmulas (P0-1) · DEDUP por código de bodega resuelto:
    # formula_items no tiene UNIQUE → dos filas del mismo material (o dos ids de
    # fórmula que mapean a la misma bodega) descontaban el DOBLE. Acumulamos por el
    # código de bodega resuelto y SUMAMOS los gramos · una sola Salida por MP.
    _acc = {}   # cod_bodega → {nombre, codigo_mp_formula, cantidad_g}
    for cod, nom, pct, g_lote in rows:
        g_total = float(g_lote or 0) * lotes
        if g_total <= 0 and cant_kg_total > 0:
            g_total = (float(pct or 0) / 100.0) * cant_kg_total * 1000.0
        if g_total <= 0:
            continue
        # resolver el id de fórmula → id de bodega (movimientos) · caso glucosamina
        cod_bodega = _resolver_material_bodega(c, cod or '', nom or '') or (cod or '')
        if not cod_bodega:
            continue
        a = _acc.setdefault(cod_bodega, {
            'codigo_mp': cod_bodega, 'codigo_mp_formula': cod or '',
            'nombre': nom or '', 'cantidad_g': 0.0,
        })
        a['cantidad_g'] += g_total
    # controla_stock: MP de fabricación propia/infinita (AGUA del lab) → la
    # producción NO la exige ni la descuenta (nunca bloquea por "no hay"). Se
    # marca con controla_stock=0 en maestro_mps (mig 218). Chequeamos tanto el
    # código de fórmula como el resuelto (cualquiera marcado → no controla).
    def _no_controla(cod_bodega, cod_formula):
        for cc in (cod_bodega, cod_formula):
            if not cc:
                continue
            try:
                r = c.execute("SELECT COALESCE(controla_stock,1) FROM maestro_mps WHERE codigo_mp=?", (cc,)).fetchone()
                if r is not None and int(r[0] or 0) == 0:
                    return True
            except Exception:
                pass
        return False
    mps = [{'codigo_mp': k, 'codigo_mp_formula': v['codigo_mp_formula'],
            'nombre': v['nombre'], 'cantidad_g': round(v['cantidad_g'], 2),
            'controla_stock': 0 if _no_controla(k, v['codigo_mp_formula']) else 1}
           for k, v in _acc.items()]
    return mps, {
        'producto': producto, 'fecha': fecha, 'lotes': lotes,
        'cantidad_kg_total': cant_kg_total, 'pid': pid,
    }


class _DescuentoError(Exception):
    """Descuento MP fallido. .codigo: str. .payload: dict."""
    def __init__(self, mensaje, codigo='ERROR', payload=None):
        super().__init__(mensaje)
        self.codigo = codigo
        self.payload = payload or {}


# Sebastian 5-may-2026 (audit zero-error dashboard): lista canonica de
# estados de lote que NO se pueden consumir en produccion. Usar UPPER()
# en comparaciones porque la DB tiene mezcla de mayusculas/capitalizadas
# (calidad.py escribe 'Cuarentena' Capitalizado, inventario.py escribe
# 'CUARENTENA' UPPERCASE). UPPER normaliza ambas variantes.
_ESTADOS_LOTE_NO_PRODUCIBLES = (
    'CUARENTENA', 'CUARENTENA_EXTENDIDA',
    'VENCIDO',
    'RECHAZADO',
    'AGOTADO',
    'BLOQUEADO',
)


def _validar_stock_para_produccion(c, mps_a_consumir):
    """Verifica que haya stock suficiente para descontar TODAS las MPs.

    Devuelve lista de faltantes. Lista vacía = OK para descontar.
    Cada faltante: {codigo_mp, nombre, requerido_g, disponible_g, falta_g}.
    Excluye lotes en CUARENTENA / CUARENTENA_EXTENDIDA / VENCIDO /
    RECHAZADO / AGOTADO / BLOQUEADO (case-insensitive).
    """
    placeholders = ','.join(['?'] * len(_ESTADOS_LOTE_NO_PRODUCIBLES))
    sql = f"""
        SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0)
        FROM movimientos
        WHERE material_id=?
          AND UPPER(COALESCE(estado_lote,'')) NOT IN ({placeholders})
    """
    # 2-jun-2026 · TRANSPARENCIA "no jala lo que hay en bodega": cuando falta una
    # MP, también medimos cuánto stock de ese MISMO código está RETENIDO en estados
    # no-producibles (CUARENTENA, etc). Caso típico: bodega muestra 600g pero
    # producción ve 17.5g porque 583g están en cuarentena sin liberar por Calidad.
    sql_retenido = f"""
        SELECT UPPER(COALESCE(estado_lote,'')) AS est,
               COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0) AS stk
        FROM movimientos
        WHERE material_id=?
          AND UPPER(COALESCE(estado_lote,'')) IN ({placeholders})
        GROUP BY UPPER(COALESCE(estado_lote,''))
        HAVING stk > 0
    """
    faltantes = []
    for mp in mps_a_consumir:
        cod = mp['codigo_mp']
        if not cod:
            continue
        # MP infinita / fabricada en casa (AGUA del lab) → nunca falta, no bloquea.
        if int(mp.get('controla_stock', 1) or 0) == 0:
            continue
        params = (cod,) + _ESTADOS_LOTE_NO_PRODUCIBLES
        r = c.execute(sql, params).fetchone()
        disp = float(r[0] or 0)
        if disp + 0.01 < float(mp['cantidad_g']):
            # desglose de stock retenido por estado (cuarentena/vencido/...)
            retenido = {}
            ret_total = 0.0
            try:
                for rr in c.execute(sql_retenido, params).fetchall():
                    est = str(rr[0] or '').strip() or '(sin estado)'
                    stk = float(rr[1] or 0)
                    retenido[est] = round(stk, 2)
                    ret_total += stk
            except Exception:
                pass
            falt = {
                'codigo_mp': cod,
                'nombre': mp['nombre'],
                'codigo_mp_formula': mp.get('codigo_mp_formula') or cod,
                'requerido_g': mp['cantidad_g'],
                'disponible_g': round(disp, 2),
                'falta_g': round(mp['cantidad_g'] - disp, 2),
                'retenido_g': round(ret_total, 2),
                'retenido_por_estado': retenido,
            }
            faltantes.append(falt)

    # FIX 1-jun-2026 · PISTA de MP duplicada: si una MP falta pero hay OTRA en bodega
    # con stock y nombre similar (comparte tokens · caso 'N-acetil glucosamina' vs
    # 'N-Acetyl Glucosamina'), lo señalamos para que el usuario unifique/cree el puente.
    # Es solo diagnóstico (no se usa para descontar · ahí mandan id/bridge/nombre exacto).
    if faltantes:
        try:
            _stocked = []  # [(codigo, nombre, stock_g)] con stock disponible
            _ph = ','.join(['?'] * len(_ESTADOS_LOTE_NO_PRODUCIBLES))
            for rr in c.execute(f"""
                SELECT mm.codigo_mp,
                       COALESCE(mm.nombre_comercial, mm.nombre_inci, mm.codigo_mp),
                       COALESCE(SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad
                                         WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END),0) AS stk
                FROM maestro_mps mm
                JOIN movimientos m ON m.material_id = mm.codigo_mp
                WHERE UPPER(COALESCE(m.estado_lote,'')) NOT IN ({_ph})
                GROUP BY mm.codigo_mp
                HAVING stk > 0
            """, _ESTADOS_LOTE_NO_PRODUCIBLES).fetchall():
                _stocked.append((str(rr[0]).strip(), str(rr[1] or ''), float(rr[2] or 0)))
            for f in faltantes:
                _toks = {t for t in _norm_mp_name(f['nombre']).split() if len(t) >= 4}
                if not _toks:
                    continue
                _best = None
                for _cod2, _nom2, _stk2 in _stocked:
                    if _cod2 == f['codigo_mp']:
                        continue
                    if _toks & {t for t in _norm_mp_name(_nom2).split() if len(t) >= 4}:
                        if _best is None or _stk2 > _best[2]:
                            _best = (_cod2, _nom2, _stk2)
                if _best:
                    f['pista'] = {
                        'codigo_mp': _best[0], 'nombre': _best[1],
                        'stock_g': round(_best[2], 2),
                    }
        except Exception:
            pass
    return faltantes


def _descontar_mp_produccion(c, evento_id, user, forzar=False):
    """Helper compartido: claim atomico + descuenta MPs FEFO de una producción.

    Sebastian 5-may-2026 (Luis Enrique): "cuando carguen produccion de materia
    prima de una descuente". Antes el descuento estaba en /completar (paso
    final) pero los operarios no lo ejecutaban — productions quedaban en
    envasado y MPs nunca se descontaban del inventario. Ahora el descuento
    ocurre al INICIAR (cuando MP fisicamente sale de bodega).

    Args:
      c: cursor SQLite (caller controla la tx)
      evento_id: id de produccion_programada
      user: username del actor (audit)
      forzar: si True, ignora claim atomico (solo admin)

    Atomico: usa UPDATE-WHERE inventario_descontado_at='' para evitar
    descuento doble en race conditions.

    Levanta _DescuentoError si:
      - YA_DESCONTADO: producción ya descontada antes
      - SIN_STOCK: alguna MP no tiene stock suficiente
      - SIN_FORMULA: producto sin formula registrada (formula_items vacio)
      - PRODUCCION_NO_EXISTE

    Returns dict {ok, mps_descontadas:[...], inventario_descontado_at, total_g}.
    """
    from datetime import datetime as _dt
    fecha_iso = _dt.now().isoformat(timespec='seconds')

    mps_a_consumir, meta = _calcular_mp_consumo_produccion(c, evento_id)
    if meta is None:
        raise _DescuentoError('Producción no encontrada', 'PRODUCCION_NO_EXISTE')
    if not mps_a_consumir:
        # Formula vacía o sin %/g_por_lote: NO bloqueante · iniciar igual
        # con warning. NO marcamos inventario_descontado_at para permitir
        # re-intento si Tecnica completa la formula despues.
        log.warning(
            'iniciar_produccion sin formula valida: producto=%s evento_id=%s · '
            'inicio se permite pero NO se descuenta inventario',
            meta['producto'], evento_id,
        )
        return {
            'ok': True,
            'mps_descontadas': [],
            'inventario_descontado_at': None,
            'total_g': 0,
            'producto': meta['producto'],
            'sin_formula': True,
            'warning': (f"Producto '{meta['producto']}' sin formula valida "
                         f"(porcentaje y cantidad_g_por_lote ambos en 0/null). "
                         f"Inicio registrado pero NO se descontó inventario · "
                         f"revisa formula en /tecnica y usa /completar para "
                         f"descontar despues."),
        }

    # Pre-check stock ANTES del claim para abortar limpio sin tocar la fila.
    faltantes = _validar_stock_para_produccion(c, mps_a_consumir)
    if faltantes:
        raise _DescuentoError(
            f'Stock insuficiente para producir {meta["producto"]}: '
            f'{len(faltantes)} MP(s) sin stock',
            'SIN_STOCK',
            {'faltantes': faltantes, 'producto': meta['producto']},
        )

    # ATOMIC CLAIM · 19-may-2026 BUG-6 audit Planta PERFECTA:
    # Antes el path `forzar=True` (re-descontar tras revertir manual) hacía
    # UPDATE sin condición → 2 requests paralelos pasaban el claim, ambos
    # llegaban al loop de INSERT a movimientos y descontaban 2×. Ahora
    # `forzar=True` también es atómico: lee el descontado_at previo y solo
    # gana el claim si ese valor sigue intacto (compare-and-swap).
    if forzar:
        prev = c.execute(
            "SELECT COALESCE(inventario_descontado_at,'') "
            "FROM produccion_programada WHERE id=?",
            (evento_id,),
        ).fetchone()
        prev_at = (prev[0] if prev else '') or ''
        cur = c.execute(
            "UPDATE produccion_programada SET inventario_descontado_at=? "
            "WHERE id=? AND COALESCE(inventario_descontado_at,'')=?",
            (fecha_iso, evento_id, prev_at),
        )
    else:
        cur = c.execute(
            "UPDATE produccion_programada SET inventario_descontado_at=? "
            "WHERE id=? AND COALESCE(inventario_descontado_at,'')=''",
            (fecha_iso, evento_id),
        )
    if cur.rowcount == 0:
        actual = c.execute(
            "SELECT inventario_descontado_at FROM produccion_programada WHERE id=?",
            (evento_id,),
        ).fetchone()
        actual_at = (actual[0] if actual else '') or ''
        # En `forzar`, rowcount=0 puede significar: ya cambió desde que
        # leímos (otro request paralelo lo descontó) · reportamos race.
        codigo_err = 'YA_DESCONTADO_RACE' if forzar else 'YA_DESCONTADO'
        raise _DescuentoError(
            f'Inventario ya fue descontado el {actual_at}',
            codigo_err,
            {'inventario_descontado_at': actual_at},
        )

    # FEFO real por lote
    obs_base = (f"Producción INICIADA: {meta['producto']} — {meta['fecha']} — "
                f"{meta['lotes']} lote(s) × {meta['cantidad_kg_total']:.0f}kg")
    descontados = []
    for mp in mps_a_consumir:
        # MP infinita / fabricada en casa (AGUA del lab) → no se descuenta del
        # kardex (no se compra ni se controla). Se registra en el legajo sin Salida.
        if int(mp.get('controla_stock', 1) or 0) == 0:
            mp['distribucion_fefo'] = []
            mp['no_controla_stock'] = True
            descontados.append(mp)
            continue
        distrib = _distribuir_fefo(c, mp['codigo_mp'], mp['cantidad_g'])
        mp['distribucion_fefo'] = []
        for d in distrib:
            lote_fragment = d['lote'] or '(sin lote — stock legacy)'
            obs_mp = (obs_base + f" | FEFO lote: {lote_fragment}" +
                       (f" (vence {d['fecha_vencimiento']})"
                        if d['fecha_vencimiento'] else ""))
            try:
                c.execute("""
                    INSERT INTO movimientos
                      (material_id, material_nombre, cantidad, tipo, fecha,
                       observaciones, operador, lote, produccion_id)
                    VALUES (?, ?, ?, 'Salida', ?, ?, ?, ?, ?)
                """, (mp['codigo_mp'], mp['nombre'], d['cantidad'],
                      fecha_iso, obs_mp, user, d['lote'], evento_id))
            except Exception:
                # mig 201 aún no aplicada (PG sin produccion_id) · fallback sin
                # la columna · la reversión cae al LIKE legacy hasta aplicarla.
                c.execute("""
                    INSERT INTO movimientos
                      (material_id, material_nombre, cantidad, tipo, fecha,
                       observaciones, operador, lote)
                    VALUES (?, ?, ?, 'Salida', ?, ?, ?, ?)
                """, (mp['codigo_mp'], mp['nombre'], d['cantidad'],
                      fecha_iso, obs_mp, user, d['lote']))
            mp['distribucion_fefo'].append({
                'lote': d['lote'],
                'cantidad_g': d['cantidad'],
                'fecha_vencimiento': d['fecha_vencimiento'],
                'sin_lote': d['sin_lote'],
            })
        descontados.append(mp)

    return {
        'ok': True,
        'mps_descontadas': descontados,
        'inventario_descontado_at': fecha_iso,
        'total_g': round(sum(m['cantidad_g'] for m in descontados), 2),
        'producto': meta['producto'],
    }


def _distribuir_fefo(c, codigo_mp, cantidad_a_descontar):
    """Distribuye una cantidad a descontar entre lotes activos siguiendo FEFO
    (First-Expired-First-Out): consume primero del lote con fecha_vencimiento
    más cercana, luego del siguiente, etc.

    Sebastian (29-abr-2026): "FEFO perfecto". Antes el descuento era global
    (anotaba el lote sugerido en obs pero no era estricto). Ahora el descuento
    es POR LOTE: cada movimiento de Salida lleva el codigo del lote del que
    realmente se está consumiendo.

    Returns: lista de dicts [{lote, cantidad, fecha_vencimiento, sin_lote}, ...]
    Si la suma de stock por lotes < cantidad_a_descontar, el remainder se
    devuelve con sin_lote=True (consumo de stock sin trazabilidad de lote —
    suele indicar drift del inventario o entradas históricas sin lote).
    """
    if cantidad_a_descontar <= 0:
        return []

    # Stock disponible por lote: SUM(Entradas con lote) - SUM(Salidas con lote)
    # Sebastian 5-may-2026 (audit zero-error): ANTES solo excluia
    # ('Vencido','Bloqueado','Rechazado') case-sensitive — NO excluia
    # CUARENTENA y NO matcheaba 'VENCIDO'/'CUARENTENA' en mayusculas.
    # Riesgo: FEFO podia consumir lotes en cuarentena o vencidos.
    # Fix: UPPER() + lista canonica _ESTADOS_LOTE_NO_PRODUCIBLES.
    placeholders = ','.join(['?'] * len(_ESTADOS_LOTE_NO_PRODUCIBLES))
    # Sebastian 8-may-2026 FEFO zero-error: ANTES GROUP BY (lote, fv)
    # generaba grupos separados para movs con fv NULL vs fv real del
    # mismo lote · ORDER BY podía elegir el grupo NULL primero.
    # Fix: agrupar SOLO por lote y tomar la fv real desde la Entrada
    # (que es donde el proveedor la declara · MAX ignora NULLs).
    # Subconsulta: los alias fv_real/stock_lote no se pueden usar en HAVING
    # ni dentro de expresiones de ORDER BY en Postgres · envolviéndolos en
    # un FROM (...) se vuelven columnas reales (portátil SQLite/Postgres).
    sql = f"""
        SELECT lote, fv_real, stock_lote FROM (
            SELECT lote,
                   MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) AS fv_real,
                   SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stock_lote
            FROM movimientos
            WHERE material_id = ?
              AND COALESCE(lote, '') != ''
              AND UPPER(COALESCE(estado_lote, '')) NOT IN ({placeholders})
            GROUP BY lote
        ) sub
        WHERE stock_lote > 0
        ORDER BY COALESCE(fv_real, '9999-12-31') ASC,
                 lote ASC
    """
    params = (codigo_mp,) + _ESTADOS_LOTE_NO_PRODUCIBLES
    rows = c.execute(sql, params).fetchall()

    distribucion = []
    restante = float(cantidad_a_descontar)
    for lote, fv, stock_lote in rows:
        if restante <= 0:
            break
        toma = min(float(stock_lote), restante)
        distribucion.append({
            'lote': lote,
            'cantidad': round(toma, 2),
            'fecha_vencimiento': fv,
            'sin_lote': False,
        })
        restante -= toma

    # FIX-B3 13-may-2026: si aún queda cantidad por descontar, antes
    # SIEMPRE se agregaba sin_lote=True asumiendo que era stock legacy
    # (entradas históricas sin trazabilidad). Pero esto ocultaba race
    # conditions: 2 producciones paralelas mismo MP, la 2da encontraba
    # solo 60g, faltaban 40g, se agregaba sin_lote silenciosamente
    # creando STOCK FANTASMA NEGATIVO oculto por max(...,0) en _get_mp_stock.
    # Ahora chequeamos primero si hay stock legacy real (movs con lote='');
    # solo si lo hay y cubre el restante, permitimos el sin_lote. Sino,
    # raise _DescuentoError para que la operación haga ROLLBACK limpio.
    if restante > 0.01:
        # Audit 3-jun · el stock legacy (lote='') DEBE excluir estados no-producibles
        # (cuarentena/vencido/bloqueado/etc.), igual que el FEFO por lote y la
        # validación · si no, descontaría stock retenido por Calidad.
        _ph_leg = ','.join(['?'] * len(_ESTADOS_LOTE_NO_PRODUCIBLES))
        legacy_row = c.execute(f"""
            SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0)
            FROM movimientos
            WHERE material_id = ? AND COALESCE(lote,'') = ''
              AND UPPER(COALESCE(estado_lote,'')) NOT IN ({_ph_leg})
        """, (codigo_mp,) + _ESTADOS_LOTE_NO_PRODUCIBLES).fetchone()
        legacy_stock = float(legacy_row[0] or 0)
        if legacy_stock + 0.01 < restante:
            # Race condition o drift · NO insertar sin_lote silenciosamente
            raise _DescuentoError(
                f"Stock real insuficiente para {codigo_mp}: faltan "
                f"{restante:.2f}g (stock legacy sin lote disponible: "
                f"{legacy_stock:.2f}g). Posible race condition o drift.",
                'SIN_STOCK',
                {
                    'codigo_mp': codigo_mp,
                    'faltante_g': round(restante, 2),
                    'legacy_g': round(legacy_stock, 2),
                    'razon': 'race_o_drift_post_validacion',
                },
            )
        # Caso válido: hay stock legacy suficiente
        distribucion.append({
            'lote': None,
            'cantidad': round(restante, 2),
            'fecha_vencimiento': None,
            'sin_lote': True,
        })

    return distribucion


# ── Helper: descuento MEE al terminar envasado ────────────────────────────
# Sebastian (1-may-2026): "en produccion dicen envasado, para que coloquen
# cuanto fue y de alli mismo descuenta automaticamente envases y demas".
# Hoy el descuento ocurre al COMPLETAR producción usando cantidad planificada.
# Esta función mueve el descuento al TERMINAR envasado, usando la cantidad
# REAL envasada (proporcional al plan).
#
# Filosofía:
#   - Se descuentan SOLO los componentes que se usan físicamente al envasar:
#     envase_primario, envase_secundario, tapa, caja_exterior, etiquetas.
#   - Serigrafía/tampografía se descuentan en el evento de decoración (D-20),
#     no aquí (su movimiento se hace cuando el operario saca de bodega).
#   - Si unidades_envasadas < unidades_planeadas → descuenta proporcional
#     (la merma queda implícita: stock MEE = stock_actual - real_consumido).
#   - Marca consumido_at + consumido_contexto='envasado' en checklist para
#     que prog_completar_evento NO vuelva a descontar el mismo ítem.
TIPOS_MEE_AL_ENVASAR = (
    'envase_primario', 'envase_secundario', 'envase',
    'tapa',
    'caja_exterior', 'caja',
    'etiqueta_frontal', 'etiqueta_posterior', 'etiqueta_lateral', 'etiqueta',
    'etiqueta_adhesiva',
)


def _descontar_mee_envasado(c, produccion_id, lote, unidades_envasadas,
                             unidades_planeadas, user):
    """Descuenta MEE de envase/tapa/etiqueta al terminar envasado.

    FIX B2B multi-envase 24-may-2026: si hay aportes B2B con envase
    específico (pedidos_b2b_lote.envase_codigo), el descuento se SPLIT:
    - Tapa/etiqueta del checklist → siempre default (las usa todos clientes).
    - ENVASE del checklist (item_tipo='envase') → descuento default solo
      por las uds que NO van a envase B2B custom.
    - Por cada envase B2B custom → descuento separado en movimientos_mee
      a razón de 1 envase por unidad B2B (ratio aplicado).

    Args:
      c: cursor SQLite
      produccion_id: int — produccion_programada.id (lote padre)
      lote: str — lote del envasado (para batch_ref en movimientos)
      unidades_envasadas: int — cantidad REAL envasada (registrada por operario)
      unidades_planeadas: int — cantidad planeada de envasado (de produccion_envasado)
      user: str — operador que termina el envasado

    Returns: lista de dicts con descontados, o [] si no aplica.
    """
    if not produccion_id or not unidades_envasadas or unidades_envasadas <= 0:
        return []

    # Ratio para cantidades 1:1 con unidades. Si planeadas no está, asume 100%.
    if unidades_planeadas and unidades_planeadas > 0:
        ratio = float(unidades_envasadas) / float(unidades_planeadas)
    else:
        ratio = 1.0

    # FEATURE B2B multi-envase · leer aportes con envase custom para split.
    # Cada aporte B2B con envase_codigo no vacío resta del envase default.
    b2b_envases_custom = []  # list of (envase_codigo, uds_real_b2b, cliente)
    uds_b2b_custom_total = 0
    try:
        for ar in c.execute(
            """SELECT pbl.envase_codigo, pbl.unidades_aporte,
                      COALESCE(pbl.cliente_nombre, '')
               FROM pedidos_b2b_lote pbl
               WHERE pbl.lote_produccion_id = ?
                 AND COALESCE(pbl.envase_codigo, '') != ''""",
            (produccion_id,),
        ).fetchall():
            env_cod = (ar[0] or '').strip().upper()
            uds_plan_b2b = int(ar[1] or 0)
            if not env_cod or uds_plan_b2b <= 0:
                continue
            uds_real_b2b = int(round(uds_plan_b2b * ratio))
            if uds_real_b2b <= 0:
                continue
            b2b_envases_custom.append((env_cod, uds_real_b2b, ar[2] or ''))
            uds_b2b_custom_total += uds_real_b2b
    except Exception:
        pass  # mig 171/172 no aplicada · split deshabilitado, descuenta default

    # Construir whitelist de tipos en SQL
    placeholders = ','.join(['?'] * len(TIPOS_MEE_AL_ENVASAR))
    sql = f"""
        SELECT id, mee_codigo_asignado, descripcion, cantidad_unidades, item_tipo
        FROM produccion_checklist
        WHERE produccion_id = ?
          AND COALESCE(mee_codigo_asignado, '') != ''
          AND COALESCE(cantidad_unidades, 0) > 0
          AND COALESCE(consumido_at, '') = ''
          AND LOWER(COALESCE(item_tipo, '')) IN ({placeholders})
    """
    params = (produccion_id,) + tuple(t.lower() for t in TIPOS_MEE_AL_ENVASAR)
    try:
        items = c.execute(sql, params).fetchall()
    except sqlite3.OperationalError:
        return []  # tabla no existe → no descontar

    if not items and not b2b_envases_custom:
        return []

    fecha_iso = datetime.now().isoformat(timespec='seconds')
    obs_base = (f"Envasado terminado: produccion #{produccion_id} "
                f"lote {lote or 'sin-lote'} — {unidades_envasadas}/"
                f"{unidades_planeadas or '?'} ud (ratio {ratio:.2f})")

    # Audit zero-error 2-may-2026: usar aplicar_movimiento_mee · INSERT
    # movimiento + UPDATE stock_actual atómicos · garantiza drift=0 +
    # clamp si stock_nuevo < 0 (no permite stock fantasma negativo).
    from inventario_helpers import aplicar_movimiento_mee
    descontados = []
    for item_id, mee_cod, desc, cant_plan, tipo_item in items:
        cant_plan_f = float(cant_plan or 0)
        cant_real = round(cant_plan_f * ratio, 0)
        # Split: si es ENVASE (no tapa/etiqueta) y hay aportes B2B con
        # envase custom, restar uds B2B del descuento default.
        es_envase = (tipo_item or '').lower() in ('envase', 'frasco', 'recipiente')
        if es_envase and uds_b2b_custom_total > 0:
            cant_real = max(cant_real - uds_b2b_custom_total, 0)
        if cant_real <= 0:
            continue
        try:
            aplicar_movimiento_mee(
                c.connection, mee_cod, 'Salida', cant_real,
                observaciones=obs_base, responsable=user,
                lote_ref=str(produccion_id), batch_ref=lote or '',
            )
            c.execute("""
                UPDATE produccion_checklist
                   SET consumido_at = ?,
                       consumido_por = ?,
                       cantidad_consumida_real = ?,
                       consumido_contexto = 'envasado',
                       actualizado_at = datetime('now', '-5 hours')
                 WHERE id = ?
            """, (fecha_iso, user, cant_real, item_id))
            descontados.append({
                'codigo': mee_cod,
                'descripcion': desc or '',
                'tipo_item': tipo_item or '',
                'cantidad_planeada': cant_plan_f,
                'cantidad_real': cant_real,
                'merma': max(0, cant_plan_f - cant_real),
                'split_b2b': es_envase and uds_b2b_custom_total > 0,
            })
        except sqlite3.OperationalError as e:
            # Si maestro_mee o movimientos_mee no existe, dejamos pasar
            log.warning(f'Descuento MEE skip {mee_cod}: {e}')
            continue

    # Descuento separado de envases B2B custom · 1:1 con uds_real_b2b.
    for env_cod, uds_real_b2b, cliente in b2b_envases_custom:
        obs_b2b = (f"Envasado B2B custom: produccion #{produccion_id} "
                    f"lote {lote or 'sin-lote'} — cliente {cliente or '?'} · "
                    f"{uds_real_b2b} ud envase {env_cod}")
        try:
            aplicar_movimiento_mee(
                c.connection, env_cod, 'Salida', uds_real_b2b,
                observaciones=obs_b2b, responsable=user,
                lote_ref=str(produccion_id), batch_ref=lote or '',
            )
            descontados.append({
                'codigo': env_cod,
                'descripcion': f'Envase B2B custom · {cliente}',
                'tipo_item': 'envase_b2b',
                'cantidad_planeada': uds_real_b2b,
                'cantidad_real': uds_real_b2b,
                'merma': 0,
                'cliente_b2b': cliente,
            })
        except sqlite3.OperationalError as e:
            log.warning(f'Descuento MEE B2B custom skip {env_cod}: {e}')
            continue
        except Exception as e:
            log.warning(f'Descuento MEE B2B custom err {env_cod}: {e}')
            continue

    return descontados


@bp.route('/api/programacion/programar/<int:evento_id>/completar', methods=['POST'])
def prog_completar_evento(evento_id):
    """Marca una produccion como completada Y descuenta inventario.

    Sebastian (29-abr-2026): "que todo descuente que el inventario este
    perfecto". Antes solo cambiaba estado — ahora:
      1. Calcula consumo de MPs desde formula_items: cantidad_g_por_lote * lotes
         (o si no hay cantidad_g_por_lote, usa porcentaje * cantidad_kg * 1000).
      2. Calcula consumo de MEEs desde produccion_checklist (items con
         mee_codigo_asignado en estado verificado_ok / recibido / listo).
         · Excluye items con consumido_at != NULL (ya descontados al envasar).
      3. Inserta movimientos de Salida por cada MP.
      4. Inserta movimientos_mee + actualiza maestro_mee.stock_actual por MEE.
      5. UPDATE produccion_programada estado='completado',
         inventario_descontado_at=NOW.
      6. Idempotente: si ya tiene inventario_descontado_at, no descontar 2x.

    Body opcional: {forzar_redescuento: true} para casos de emergencia
    (ej. el flag quedó stale por bug). Solo admin.
    Body opcional: {dry_run: true} para preview sin escribir nada.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.get_json() or {}
    forzar = bool(d.get('forzar_redescuento'))
    dry_run = bool(d.get('dry_run'))
    if forzar and user not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin puede forzar re-descuento'}), 403

    conn = get_db(); c = conn.cursor()
    # BUG-1 fix · 19-may-2026: solo operario asignado / jefe / admin
    # puede completar y descontar MPs/MEEs de la producción.
    ok_caller, err_caller = _caller_puede_operar_produccion(c, user, evento_id)
    if not ok_caller:
        return err_caller
    prod = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
               pp.estado, COALESCE(pp.inventario_descontado_at,'') as descontado_at,
               COALESCE(pp.cantidad_kg, 0)         as cantidad_kg_explicita,
               COALESCE(fh.lote_size_kg, 0)        as lote_kg_formula
        FROM produccion_programada pp
        LEFT JOIN formula_headers fh
               ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
        WHERE pp.id=?
    """, (evento_id,)).fetchone()
    if not prod:
        return jsonify({'error': 'Producción no encontrada'}), 404
    pid, producto, fecha, lotes, estado, descontado_at, cant_kg_exp, lote_kg = prod
    lotes = int(lotes or 1)
    cant_kg_total = float(cant_kg_exp or 0) or (lotes * float(lote_kg or 0))

    # Sebastian 5-may-2026 (Luis Enrique): si ya descontó al INICIAR (flujo
    # nuevo · auto descuenta MP en /iniciar), permitimos que /completar siga
    # corriendo solo para MEE + estado=completado · NO devolvemos 409.
    # Solo bloqueamos si forzar=False AND aún hay que re-descontar MP — eso
    # solo ocurre si admin pide forzar (chequeado abajo). El skip de MP se
    # hace en el bloque ATOMIC CLAIM (rowcount=0 ya no aborta · solo skip).
    mp_ya_descontado = bool(descontado_at) and not forzar

    # ── 1. Calcular MPs a consumir ─────────────────────────────────────
    # FIX 1-jun-2026 audit MP/fórmulas (P0) · ANTES recalculaba inline con el
    # material_id de FÓRMULA crudo (sin resolver a bodega) → reintroducía 'Hay 0g'
    # al completar (caso glucosamina) y duplicaba el cálculo de iniciar. Ahora usa
    # el helper canónico _calcular_mp_consumo_produccion (resuelve fórmula→bodega +
    # dedup) · UN solo punto de cálculo para iniciar y completar.
    try:
        mps_a_consumir, _meta_mp = _calcular_mp_consumo_produccion(c, evento_id)
        if mps_a_consumir is None:
            mps_a_consumir = []
    except Exception as e:
        return jsonify({'error': f'Error calculando MPs: {e}'}), 500

    # ── 2. Calcular MEEs a consumir desde checklist (solo recibidos/verificados
    #       y NO consumidos previamente al envasar) ──
    # Sebastian (1-may-2026): los items envase/tapa/etiqueta se descuentan al
    # terminar envasado con cantidad REAL (helper _descontar_mee_envasado).
    # Aquí solo descontamos lo que NO se consumió aún (serigrafía/tampografía/
    # caja_master si vienen al final).
    mees_a_consumir = []
    mee_item_ids = []  # ids del checklist para marcar consumido_at después
    try:
        crows = c.execute("""
            SELECT id, mee_codigo_asignado, descripcion, cantidad_unidades, item_tipo
            FROM produccion_checklist
            WHERE produccion_id = ?
              AND COALESCE(mee_codigo_asignado,'') != ''
              AND COALESCE(cantidad_unidades, 0) > 0
              AND COALESCE(consumido_at, '') = ''
              AND estado IN ('verificado_ok','recibido','listo')
        """, (pid,)).fetchall()
        for item_id, mee_cod, desc, cant_ud, tipo_item in crows:
            mees_a_consumir.append({
                'item_id': item_id,
                'codigo': mee_cod, 'descripcion': desc or '',
                'cantidad_unidades': int(cant_ud or 0),
                'tipo_item': tipo_item or '',
            })
            mee_item_ids.append(item_id)
    except Exception:
        # Si la tabla checklist no existe o falla, seguimos sin MEEs (no bloqueamos MPs)
        pass

    # ── 3. Si dry_run, devolver preview sin escribir ───────────────────
    if dry_run:
        return jsonify({
            'ok': True,
            'dry_run': True,
            'producto': producto,
            'fecha': fecha,
            'lotes': lotes,
            'cantidad_kg_total': cant_kg_total,
            'mps_a_descontar': mps_a_consumir,
            'mees_a_descontar': mees_a_consumir,
            'total_mps': len(mps_a_consumir),
            'total_mees': len(mees_a_consumir),
            'total_g_mps': round(sum(m['cantidad_g'] for m in mps_a_consumir), 2),
            'total_unidades_mees': sum(m['cantidad_unidades'] for m in mees_a_consumir),
        })

    # ── 4. ESCRIBIR descuentos (transaccional) ─────────────────────────
    obs_base = f"Producción COMPLETADA: {producto} — {fecha} — {lotes} lote(s) × {cant_kg_total:.0f}kg"
    descontados_mps = []
    descontados_mees = []
    fecha_iso = __import__('datetime').datetime.now().isoformat(timespec='seconds')
    try:
        # ATOMIC CLAIM (audit zero-error 2-may-2026 · CERO SESGO).
        # Sebastian 5-may-2026: skip si ya descontó al iniciar (no devolvemos
        # 409 · solo procesamos MEE + cierre).
        if mp_ya_descontado:
            claim_rowcount = 0  # ya descontado, skip MP loop
        elif forzar:
            # CAS atómico (fix 28-may) · antes UPDATE incondicional → dos
            # /completar?forzar paralelos pasaban ambos y doble-descontaban MP.
            # Ahora solo gana el request que coincide con el descontado_at
            # previo leído · el segundo cae en el 409 YA_DESCONTADO_RACE.
            claim_cur = c.execute("""
                UPDATE produccion_programada
                   SET inventario_descontado_at = ?
                 WHERE id = ?
                   AND COALESCE(inventario_descontado_at, '') = COALESCE(?, '')
            """, (fecha_iso, evento_id, descontado_at))
            claim_rowcount = claim_cur.rowcount
        else:
            claim_cur = c.execute("""
                UPDATE produccion_programada
                   SET inventario_descontado_at = ?
                 WHERE id = ?
                   AND COALESCE(inventario_descontado_at, '') = ''
            """, (fecha_iso, evento_id))
            claim_rowcount = claim_cur.rowcount
        if claim_rowcount == 0 and not mp_ya_descontado:
            # Race: otro request descontó al mismo tiempo · liberar tx + 409
            try: conn.rollback()
            except Exception: pass
            actual = c.execute(
                "SELECT inventario_descontado_at FROM produccion_programada WHERE id=?",
                (evento_id,)).fetchone()
            actual_at = actual[0] if actual else descontado_at
            return jsonify({
                'error': f'Inventario ya fue descontado el {actual_at}',
                'codigo': 'YA_DESCONTADO_RACE',
                'hint': 'Otro proceso descontó al mismo tiempo · idempotencia respetada',
                'inventario_descontado_at': actual_at,
            }), 409
        # MPs → FEFO REAL: distribuir el consumo entre lotes según fecha
        # de vencimiento (más cercano primero). Cada lote genera su propio
        # movimiento de Salida con cantidad específica. Si la suma de stock
        # por lote no alcanza, el remainder se registra sin lote (legacy).
        # Sebastian 5-may-2026: skip si ya descontó al iniciar.
        if not mp_ya_descontado:
            for mp in mps_a_consumir:
                distrib = _distribuir_fefo(c, mp['codigo_mp'], mp['cantidad_g'])
                mp['distribucion_fefo'] = []
                for d in distrib:
                    lote_fragment = d['lote'] or '(sin lote — stock legacy)'
                    obs_mp = (obs_base +
                              f" | FEFO lote: {lote_fragment}" +
                              (f" (vence {d['fecha_vencimiento']})" if d['fecha_vencimiento'] else ""))
                    try:
                        c.execute("""
                            INSERT INTO movimientos
                              (material_id, material_nombre, cantidad, tipo, fecha,
                               observaciones, operador, lote, produccion_id)
                            VALUES (?, ?, ?, 'Salida', ?, ?, ?, ?, ?)
                        """, (mp['codigo_mp'], mp['nombre'], d['cantidad'],
                              fecha_iso, obs_mp, user, d['lote'], evento_id))
                    except Exception:
                        # mig 201 aún no aplicada (PG) · fallback sin la columna.
                        c.execute("""
                            INSERT INTO movimientos
                              (material_id, material_nombre, cantidad, tipo, fecha,
                               observaciones, operador, lote)
                            VALUES (?, ?, ?, 'Salida', ?, ?, ?, ?)
                        """, (mp['codigo_mp'], mp['nombre'], d['cantidad'],
                              fecha_iso, obs_mp, user, d['lote']))
                    mp['distribucion_fefo'].append({
                        'lote': d['lote'],
                        'cantidad_g': d['cantidad'],
                        'fecha_vencimiento': d['fecha_vencimiento'],
                        'sin_lote': d['sin_lote'],
                    })
                descontados_mps.append(mp)
        # MEEs → aplicar_movimiento_mee · drift=0 garantizado (audit zero-error)
        # lote_ref=produccion_id permite reversión limpia sin LIKE en obs.
        from inventario_helpers import aplicar_movimiento_mee as _aplicar_mee_inner
        for me in mees_a_consumir:
            try:
                _aplicar_mee_inner(
                    c.connection, me['codigo'], 'Salida', me['cantidad_unidades'],
                    observaciones=obs_base, responsable=user,
                    lote_ref=str(pid), batch_ref=str(fecha or ''),
                )
                # Marcar consumido en checklist (contexto='completar')
                if me.get('item_id'):
                    c.execute("""
                        UPDATE produccion_checklist
                           SET consumido_at = ?,
                               consumido_por = ?,
                               cantidad_consumida_real = ?,
                               consumido_contexto = 'completar',
                               actualizado_at = datetime('now', '-5 hours')
                         WHERE id = ?
                    """, (fecha_iso, user, me['cantidad_unidades'], me['item_id']))
                descontados_mees.append(me)
            except sqlite3.OperationalError as _oe:
                # BUG-9 fix · 19-may-2026 audit Planta PERFECTA: catch
                # demasiado amplio · antes silenciaba "database locked",
                # "constraint failed" y dejaba el checklist en estado
                # inconsistente (verificado_ok sin consumido_at). Ahora
                # solo continuar si es "no such table" (caso esperado de
                # MEE no migrado); cualquier otro error es real → re-raise
                # para que la transacción rollbackee.
                _msg = str(_oe).lower()
                if 'no such table' in _msg or 'no existe la' in _msg:
                    continue
                raise
        # Actualizar produccion_programada · estado + observaciones
        # (inventario_descontado_at YA fue seteado por el ATOMIC CLAIM al inicio)
        c.execute("""
            UPDATE produccion_programada
               SET estado='completado',
                   fin_real_at = COALESCE(fin_real_at, datetime('now','-5 hours')),
                   observaciones = COALESCE(observaciones,'') ||
                                   ' | INVENTARIO DESCONTADO ' || ? || ' por ' || ?
             WHERE id=?
        """, (fecha_iso, user, pid))
        # Audit log INVIMA · dispensación es operación regulada GMP/BPM
        # (Resolución 2214/2021). Trazabilidad obligatoria: quién dispensó qué
        # producción y cuándo se descontó inventario.
        try:
            audit_log(c, usuario=user, accion='COMPLETAR_PRODUCCION',
                      tabla='produccion_programada', registro_id=pid,
                      despues={'producto': producto, 'lotes': lotes,
                                'fecha_programada': fecha,
                                'inventario_descontado_at': fecha_iso})
        except Exception as e:
            log.warning('audit_log COMPLETAR_PRODUCCION fallo: %s', e)

        # Sebastián 1-may-2026: "queden limpias el mismo día". Si la
        # producción tenía área asignada → marcar sucia + crear limpieza HOY
        # con operario rotando.
        limpieza_id = None
        try:
            row = c.execute("""
                SELECT pp.area_id, ap.codigo, pp.lotes
                FROM produccion_programada pp
                  LEFT JOIN areas_planta ap ON ap.id = pp.area_id
                WHERE pp.id=?
            """, (pid,)).fetchone()
            if row and row[0]:
                area_id_, area_codigo, lotes_n = row
                # Marcar sucia
                c.execute(
                    "UPDATE areas_planta SET estado='sucia' WHERE id=? AND estado='ocupada'",
                    (area_id_,)
                )
                # Crear limpieza mismo día
                limpieza_id = _crear_limpieza_post_produccion(
                    c, area_id_, area_codigo, fecha,
                    producto, f'{lotes_n}lt' if lotes_n else '', user
                )
        except Exception as _e:
            log.warning(f'[completar] limpieza post-prod falla: {_e}')

        conn.commit()
    except Exception as e:
        try: conn.rollback()
        except Exception as _r:
            logging.getLogger('programacion').debug('rollback no aplicable: %s', _r)
        return jsonify({
            'error': f'Error descontando inventario: {e}',
            'mps_descontados_antes_de_fallar': descontados_mps,
            'mees_descontados_antes_de_fallar': descontados_mees,
        }), 500

    return jsonify({
        'ok': True,
        'id': pid,
        'producto': producto,
        'inventario_descontado_at': fecha_iso,
        'mps_descontados': descontados_mps,
        'mees_descontados': descontados_mees,
        'total_mps': len(descontados_mps),
        'total_mees': len(descontados_mees),
        'total_g_mps': round(sum(m['cantidad_g'] for m in descontados_mps), 2),
        'total_unidades_mees': sum(m['cantidad_unidades'] for m in descontados_mees),
        'limpieza_auto_id': limpieza_id,
        'mensaje': (f"{producto} marcada completada. Descontados {len(descontados_mps)} MPs y "
                    f"{len(descontados_mees)} MEEs. " +
                    (f"🧹 Limpieza programada hoy (#{limpieza_id})" if limpieza_id else ""))
    })


@bp.route('/api/programacion/programar/<int:evento_id>/revertir-completado', methods=['POST'])
def prog_revertir_completado(evento_id):
    """Revierte una produccion completada: regresa MPs y MEEs al inventario,
    cambia estado a 'programado' y limpia el flag inventario_descontado_at.

    Solo admin. Sebastian (29-abr-2026): por si Sebastian o alguien marca
    completada por error o quiere re-hacer el descuento.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403

    conn = get_db(); c = conn.cursor()
    prod = c.execute(
        "SELECT producto, fecha_programada, lotes, "
        "COALESCE(inventario_descontado_at,'') FROM produccion_programada WHERE id=?",
        (evento_id,)
    ).fetchone()
    if not prod:
        return jsonify({'error': 'Producción no encontrada'}), 404
    producto, fecha, lotes, descontado_at = prod
    if not descontado_at:
        return jsonify({'error': 'Esta producción no tenía inventario descontado'}), 400

    obs_filtro = f"Producción COMPLETADA: {producto} — {fecha}"
    # El descuento de MP puede haber ocurrido en /iniciar ("Producción
    # INICIADA: ...") o en /completar ("Producción COMPLETADA: ..."). La
    # reversión DEBE buscar ambos prefijos · si solo busca COMPLETADA, una
    # producción descontada al iniciar limpia inventario_descontado_at sin
    # devolver la MP al stock → doble descuento al re-iniciar.
    obs_filtro_ini = f"Producción INICIADA: {producto} — {fecha}"
    # BUG-10 fix · 19-may-2026 audit Planta PERFECTA: SQL LIKE trata
    # `%` y `_` como wildcards. Si producto contiene `%` literal (ej.
    # "Crema 50%"), el LIKE matchea OTRAS producciones con el mismo
    # prefijo. Escapamos %/_/\\ usando ESCAPE '\\'. También ayuda con
    # underscore: "Suero_X" → "Suero\\_X".
    def _escape_like(s):
        return (s or '').replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
    obs_filtro_esc = _escape_like(obs_filtro)
    obs_filtro_ini_esc = _escape_like(obs_filtro_ini)
    revertidos_mps = []
    revertidos_mees = []
    fecha_iso = __import__('datetime').datetime.now().isoformat(timespec='seconds')
    try:
        # Reversión de MPs: insertar movimientos de Entrada compensatorios
        # PRESERVANDO el lote (FEFO reverso — cada lote consumido vuelve
        # a su lote de origen para que el stock por-lote quede coherente).
        # Fix 28-may (mig 201) · filtrar por produccion_id EXACTO para no
        # revertir el MP de OTRA producción del mismo producto+fecha
        # (cross-reversal → inventario fantasma). Fallback al LIKE por texto
        # solo para movimientos legacy SIN produccion_id (pre-migración).
        try:
            rows = c.execute("""
                SELECT id, material_id, material_nombre, cantidad, lote, fecha_vencimiento
                FROM movimientos
                WHERE tipo='Salida'
                  AND (produccion_id = ?
                       OR (produccion_id IS NULL
                           AND (observaciones LIKE ? ESCAPE '\\'
                                OR observaciones LIKE ? ESCAPE '\\')))
            """, (evento_id, f"{obs_filtro_esc}%", f"{obs_filtro_ini_esc}%")).fetchall()
        except Exception:
            # mig 201 aún no aplicada (PG sin produccion_id) · degradar al LIKE
            # legacy (comportamiento previo · con su limitación conocida).
            rows = c.execute("""
                SELECT id, material_id, material_nombre, cantidad, lote, fecha_vencimiento
                FROM movimientos
                WHERE tipo='Salida'
                  AND (observaciones LIKE ? ESCAPE '\\'
                       OR observaciones LIKE ? ESCAPE '\\')
            """, (f"{obs_filtro_esc}%", f"{obs_filtro_ini_esc}%")).fetchall()
        for mid, cod, nom, cant, lote, fv in rows:
            c.execute("""
                INSERT INTO movimientos
                  (material_id, material_nombre, cantidad, tipo, fecha,
                   observaciones, operador, lote, fecha_vencimiento)
                VALUES (?, ?, ?, 'Entrada', ?, ?, ?, ?, ?)
            """, (cod, nom, cant, fecha_iso,
                  f"REVERSIÓN producción completada — original mov #{mid}",
                  user, lote, fv))
            revertidos_mps.append({
                'codigo_mp': cod, 'nombre': nom,
                'cantidad_g': cant, 'lote': lote,
            })

        # Reversión de MEEs: doble fuente
        #   1) movimientos por lote_ref = produccion_id (descuentos al envasar
        #      o al completar usan ese lote_ref desde 1-may-2026).
        #   2) Fallback legacy: observaciones LIKE 'Producción COMPLETADA%'
        # Sebastian (1-may-2026): el descuento puede ocurrir en envasado o en
        # completar, ambos quedan vinculados a produccion_id por lote_ref.
        try:
            # Primero por lote_ref (preciso, sin LIKE)
            rows_mee = c.execute("""
                SELECT id, mee_codigo, cantidad
                FROM movimientos_mee
                WHERE LOWER(tipo)='salida'
                  AND COALESCE(lote_ref,'') = ?
                  AND COALESCE(anulado,0) = 0
            """, (str(evento_id),)).fetchall()
            ids_revertidos = set(r[0] for r in rows_mee)
            # Luego fallback legacy (sin lote_ref) por observación
            rows_legacy = c.execute("""
                SELECT id, mee_codigo, cantidad
                FROM movimientos_mee
                WHERE LOWER(tipo)='salida'
                  AND COALESCE(lote_ref,'') = ''
                  AND observaciones LIKE ?
                  AND COALESCE(anulado,0) = 0
            """, (f"{obs_filtro}%",)).fetchall()
            for r in rows_legacy:
                if r[0] not in ids_revertidos:
                    rows_mee = list(rows_mee) + [r]
            # Audit zero-error 2-may-2026: usar aplicar_movimiento_mee para
            # la reversión · garantiza drift=0 (Entrada compensatoria atómica).
            #
            # IMPORTANTE: NO marcar el movimiento original como anulado.
            # Hacer AMBAS cosas (entrada compensatoria + anular original) era
            # un bug preexistente que provocaba drift = -cant en SUM(movs)
            # porque el calc excluye anulados pero stock_actual sí los suma.
            #
            # La entrada compensatoria YA es la "anulación lógica" · ambos
            # movimientos quedan en historial para auditoría completa
            # (el operador ve "Salida X · Reversión X" como pares en logs).
            from inventario_helpers import aplicar_movimiento_mee as _aplicar_mee_rev
            for mid, mee_cod, cant in rows_mee:
                # No revertir dos veces el mismo movimiento · si ya existe su
                # Entrada compensatoria, saltar. Evita inflar stock en un ciclo
                # completar→revertir→completar→revertir (la salida original
                # nunca se anula, así que volvería a aparecer en el query).
                if c.execute(
                    "SELECT 1 FROM movimientos_mee WHERE tipo='Entrada' "
                    "AND observaciones LIKE ? LIMIT 1",
                    (f"%mov_mee #{mid}%",),
                ).fetchone():
                    continue
                _aplicar_mee_rev(
                    c.connection, mee_cod, 'Entrada', cant,
                    observaciones=f"REVERSIÓN producción completada — original mov_mee #{mid}",
                    responsable=user, lote_ref=str(evento_id),
                )
                revertidos_mees.append({'codigo': mee_cod, 'cantidad_unidades': cant})
            # Resetear flags consumido_at en el checklist para esta producción
            c.execute("""
                UPDATE produccion_checklist
                   SET consumido_at = NULL,
                       consumido_por = '',
                       cantidad_consumida_real = 0,
                       consumido_contexto = ''
                 WHERE produccion_id = ?
            """, (evento_id,))
        except sqlite3.OperationalError:
            pass

        # Limpiar flag y estado
        c.execute("""
            UPDATE produccion_programada
               SET estado='programado',
                   inventario_descontado_at=NULL,
                   observaciones = COALESCE(observaciones,'') ||
                                   ' | REVERTIDO ' || ? || ' por ' || ?
             WHERE id=?
        """, (fecha_iso, user, evento_id))
        # Audit log INVIMA · revertir descuento de inventario es operación
        # regulada (inversa de COMPLETAR_PRODUCCION). Trazabilidad obligatoria.
        try:
            audit_log(c, usuario=user, accion='REVERTIR_COMPLETADO',
                      tabla='produccion_programada', registro_id=evento_id,
                      antes={'estado': 'completado', 'inventario_descontado_at': 'set'},
                      despues={'estado': 'programado', 'inventario_descontado_at': None,
                               'mps_revertidos': len(revertidos_mps),
                               'mees_revertidos': len(revertidos_mees),
                               'fecha': fecha_iso})
        except Exception as e:
            log.warning('audit_log REVERTIR_COMPLETADO fallo: %s', e)
        conn.commit()
    except Exception as e:
        try: conn.rollback()
        except Exception as _r:
            logging.getLogger('programacion').debug('rollback no aplicable: %s', _r)
        return jsonify({'error': f'Error revirtiendo: {e}'}), 500

    return jsonify({
        'ok': True,
        'id': evento_id,
        'mps_revertidos': revertidos_mps,
        'mees_revertidos': revertidos_mees,
        'mensaje': f"Producción revertida. {len(revertidos_mps)} MPs y {len(revertidos_mees)} MEEs regresados al inventario."
    })


@bp.route('/api/inventario/ajuste-manual', methods=['POST'])
def inventario_ajuste_manual():
    """Registra un ajuste manual de inventario con razón obligatoria.

    Sebastian (29-abr-2026): "que se mantenga, sea perfecta". Para los
    casos legítimos donde Sebastián/Daniela necesitan corregir stock
    (conteo físico, merma, robo, daño, etc.) — TODO ajuste queda en
    movimientos con observaciones específicas para auditoría.

    Body: {
      tipo_material: 'MP' | 'MEE',
      codigo: 'MP00001',
      cantidad: 1500,        # positivo siempre — el signo lo da motivo
      motivo: 'conteo_fisico' | 'merma' | 'robo' | 'daño' | 'correccion' | 'otro',
      direccion: 'sumar' | 'restar',
      observaciones: 'detalle libre obligatorio'
    }
    Solo admin.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin puede ajustar inventario'}), 403
    d = request.get_json() or {}
    tipo_mat = (d.get('tipo_material') or '').upper().strip()
    codigo = (d.get('codigo') or '').strip()
    cantidad = abs(float(d.get('cantidad') or 0))
    motivo = (d.get('motivo') or '').strip().lower()
    direccion = (d.get('direccion') or '').strip().lower()
    obs = (d.get('observaciones') or '').strip()

    MOTIVOS_OK = {'conteo_fisico', 'merma', 'robo', 'daño', 'dano',
                  'correccion', 'corrección', 'otro'}
    if tipo_mat not in ('MP', 'MEE'):
        return jsonify({'error': 'tipo_material debe ser MP o MEE'}), 400
    if not codigo:
        return jsonify({'error': 'codigo requerido'}), 400
    if cantidad <= 0:
        return jsonify({'error': 'cantidad debe ser > 0'}), 400
    if motivo not in MOTIVOS_OK:
        return jsonify({'error': f'motivo invalido — usar uno de {MOTIVOS_OK}'}), 400
    if direccion not in ('sumar', 'restar'):
        return jsonify({'error': "direccion debe ser 'sumar' o 'restar'"}), 400
    if not obs or len(obs) < 10:
        return jsonify({'error': 'observaciones obligatorias (min 10 chars)'}), 400

    conn = get_db(); c = conn.cursor()
    fecha_iso = __import__('datetime').datetime.now().isoformat(timespec='seconds')
    obs_full = f"AJUSTE_MANUAL[{motivo}]: {obs} (por {user})"

    if tipo_mat == 'MP':
        # Verificar que el MP existe
        mp = c.execute(
            "SELECT codigo_mp, nombre_inci FROM maestro_mps WHERE codigo_mp=?",
            (codigo,)
        ).fetchone()
        if not mp:
            return jsonify({'error': f'MP {codigo} no encontrado'}), 404
        tipo_mov = 'Entrada' if direccion == 'sumar' else 'Salida'
        c.execute("""
            INSERT INTO movimientos
              (material_id, material_nombre, cantidad, tipo, fecha,
               observaciones, operador)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (codigo, mp[1] or '', cantidad, tipo_mov, fecha_iso, obs_full, user))
    else:
        # MEE: actualizar maestro_mee + insertar movimientos_mee
        mee = c.execute(
            "SELECT codigo, descripcion, COALESCE(stock_actual,0) "
            "FROM maestro_mee WHERE codigo=?", (codigo,)
        ).fetchone()
        if not mee:
            return jsonify({'error': f'MEE {codigo} no encontrado'}), 404
        delta = cantidad if direccion == 'sumar' else -cantidad
        nuevo_stock = max(0, float(mee[2]) + delta)
        c.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?",
                  (nuevo_stock, codigo))
        c.execute("""
            INSERT INTO movimientos_mee
              (mee_codigo, tipo, cantidad, observaciones, responsable, fecha)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (codigo, 'Entrada' if direccion == 'sumar' else 'Salida',
              cantidad, obs_full, user, fecha_iso))
    conn.commit()
    return jsonify({
        'ok': True,
        'tipo_material': tipo_mat,
        'codigo': codigo,
        'cantidad': cantidad,
        'direccion': direccion,
        'motivo': motivo,
        'mensaje': f'Ajuste {direccion} {cantidad} a {codigo} registrado.'
    })


@bp.route('/api/programacion/debug-calendario')
def prog_debug_calendario():
    """Show raw calendar events and how they match to products."""
    conn = get_db()
    cal  = _fetch_calendar_events(days_ahead=120)
    conn2 = get_db()
    formulas = _get_formulas(conn2)
    products_with_formula = list(formulas.keys())

    import re as _re_dbg
    _stop = {'DE','DEL','LA','EL','LOS','LAS','CON','MAS','PARA',
             'PRODUCCION','PRODUCCIÓN','LOTE','ESPAGIRIA','ANIMUS'}

    def score(t, p):
        tu, pu = t.upper(), p.upper()
        if pu in tu: return 100
        tw = set(w for w in _re_dbg.split(r'[^A-Z0-9]+', tu) if len(w)>2)
        pw = set(w for w in _re_dbg.split(r'[^A-Z0-9]+', pu) if len(w)>2) - _stop
        if not pw: return 0
        return int(100*len(pw&tw)/len(pw))

    # Load sku map
    import re as _re_d
    _NOT_SKU = {'FAB','QC','CON','SIN','MICRO','ENVASADO','ACONDICIONAMIENTO',
                'FABRICACION','FABRICACIÓN','LANZAMIENTO','KG','MES','DIAS','ML'}
    sku_map = {}
    try:
        for r in conn.execute("SELECT UPPER(sku), producto_nombre FROM sku_producto_map WHERE activo=1").fetchall():
            sku_map[r[0]] = r[1]
    except Exception:
        pass

    def _skus_d(titulo):
        tokens = _re_d.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', titulo.upper())
        return [t for t in tokens if t not in _NOT_SKU]

    sku_matches = {}
    matches = []
    for ev in cal.get('events', []):
        titulo = ev.get('titulo','')
        fecha  = ev.get('fecha','')
        skus   = _skus_d(titulo)
        matched_prod = None
        for sku in skus:
            pn = sku_map.get(sku)
            if pn and pn in products_with_formula:
                matched_prod = pn
                if pn not in sku_matches or fecha < sku_matches[pn]:
                    sku_matches[pn] = fecha
                break
        matches.append({
            'evento': titulo, 'fecha': fecha,
            'skus_extracted': skus,
            'matched_product': matched_prod
        })

    return jsonify({
        'source': cal.get('source','?'),
        'error':  cal.get('error'),
        'total_events': len(cal.get('events',[])),
        'gcal_ical_url_set': bool(GCAL_ICAL_URL),
        'google_api_key_set': bool(GOOGLE_API_KEY),
        'sku_matches_summary': sku_matches,
        'unmatched_events': [m for m in matches if not m['matched_product']],
        'matched_events': [m for m in matches if m['matched_product']]
    })

@bp.route('/api/programacion/calendario')
def prog_calendario():
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    days = int(request.args.get('days', 90))
    return jsonify(_fetch_calendar_events(days_ahead=days))


@bp.route('/api/programacion/stock-60d')
def prog_stock_60d():
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()
    vel_data = _shopify_velocity(conn, days=60)
    mp_stock = _get_mp_stock(conn)
    formulas = _get_formulas(conn)
    cal = _fetch_calendar_events(days_ahead=90)
    china_mps_set = _get_china_mps(conn)
    projection, _ = _project_stock(conn, vel_data['prod_velocity'], formulas, mp_stock, cal.get('events', []), china_mps=china_mps_set)
    return jsonify({'proyeccion': projection, 'generado_en': datetime.now().isoformat()})


@bp.route('/api/programacion/alertas')
def prog_alertas():
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()
    vel_data = _shopify_velocity(conn, days=60)
    mp_stock = _get_mp_stock(conn)
    formulas = _get_formulas(conn)
    cal = _fetch_calendar_events(days_ahead=90)
    china_mps_set = _get_china_mps(conn)
    _, alerts = _project_stock(conn, vel_data['prod_velocity'], formulas, mp_stock, cal.get('events', []), china_mps=china_mps_set)
    return jsonify({'alertas': alerts, 'n_alertas': len(alerts)})


@bp.route('/api/programacion/productos')
def prog_productos():
    """Lista de productos con fórmulas — usada por selectores en Envasado/Acond."""
    conn = get_db()
    rows = conn.execute(
        "SELECT producto_nombre, lote_size_kg FROM formula_headers ORDER BY producto_nombre"
    ).fetchall()
    return jsonify({
        'formulas': [{'nombre': r[0], 'lote_size_kg': r[1]} for r in rows],
        'count': len(rows),
    })


@bp.route('/api/programacion/regenerar-oc', methods=['POST'])
def prog_regenerar_oc():
    """Borra solicitudes auto-generadas Pendientes + sus OCs Borrador y
    vuelve a llamar al generador con los datos ACTUALES de Programación.

    Útil cuando el cálculo cambió (más producciones programadas, nuevos
    déficits) y las solicitudes viejas tienen cantidades obsoletas.

    NO toca solicitudes en estado Aprobada/Pagada (esas se respetan como
    histórico). Solo borra Pendientes auto-generadas (observaciones LIKE
    '%Centro Programación%' o '%Centro Programacion%' o '%Auto-generada%').
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401

    conn = get_db()
    c = conn.cursor()
    deleted = {'solicitudes': 0, 'ordenes_compra': 0, 'items': 0}

    try:
        # 1. Listar solicitudes auto-generadas Pendientes y sus OCs
        rows = c.execute("""
            SELECT numero, numero_oc FROM solicitudes_compra
            WHERE estado = 'Pendiente'
              AND categoria IN ('Materia Prima', 'MP')
              AND (observaciones LIKE '%Centro Programación%'
                   OR observaciones LIKE '%Centro Programacion%'
                   OR observaciones LIKE '%Auto-generada%')
        """).fetchall()
        nums_sol = [r[0] for r in rows]
        nums_oc = [r[1] for r in rows if r[1]]

        if nums_sol:
            ph_sol = ','.join('?' * len(nums_sol))
            try:
                d = c.execute(f"DELETE FROM solicitudes_compra_items WHERE numero IN ({ph_sol})", nums_sol)
                deleted['items'] = d.rowcount or 0
            except sqlite3.OperationalError:
                pass
            d = c.execute(f"DELETE FROM solicitudes_compra WHERE numero IN ({ph_sol})", nums_sol)
            deleted['solicitudes'] = d.rowcount or 0

        if nums_oc:
            ph_oc = ','.join('?' * len(nums_oc))
            # Solo borrar OCs en estado Borrador (las Aprobadas/Pagadas son histórico)
            ocs_borrar = [r[0] for r in c.execute(
                f"SELECT numero_oc FROM ordenes_compra WHERE numero_oc IN ({ph_oc}) AND estado='Borrador'",
                nums_oc
            ).fetchall()]
            if ocs_borrar:
                ph_b = ','.join('?' * len(ocs_borrar))
                try:
                    c.execute(f"DELETE FROM ordenes_compra_items WHERE numero_oc IN ({ph_b})", ocs_borrar)
                except sqlite3.OperationalError:
                    pass
                d = c.execute(f"DELETE FROM ordenes_compra WHERE numero_oc IN ({ph_b})", ocs_borrar)
                deleted['ordenes_compra'] = d.rowcount or 0

        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({
            'error': 'Falló el borrado de solicitudes viejas',
            'detalle': str(e),
        }), 500

    # 2. Recalcular déficit con la lógica correcta (total_g - stock una sola
    # vez). Misma fuente de verdad que /programacion para no divergir.
    mp_deficit = _compute_mp_deficit_aggregated(conn, days_ahead=90)

    if not mp_deficit:
        return jsonify({'ok': True, 'mensaje': 'Sin déficits actuales — solo se borraron viejas',
                        'borradas': deleted, 'creadas': []})

    SIN_PROV = 'Sin asignar'
    grupos = {}
    for cod, info in mp_deficit.items():
        prov = info.get('proveedor') or SIN_PROV
        grupos.setdefault(prov, []).append({
            'codigo': cod,
            'nombre': info['nombre'],
            'deficit_g': info['deficit_g'],
            'productos': info['productos'],
        })

    from datetime import datetime as _dt
    year = _dt.now().strftime('%Y')
    last_n = conn.execute(
        "SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) FROM solicitudes_compra WHERE numero LIKE ?",
        (f"SOL-{year}-%",)
    ).fetchone()[0] or 0
    n_sol = last_n
    user = session.get('compras_user', 'Sistema')

    creadas = []
    proveedores_ordenados = sorted(grupos.keys(), key=lambda p: (p == SIN_PROV, p.lower()))

    try:
        for prov in proveedores_ordenados:
            items = grupos[prov]
            if not items:
                continue
            n_sol += 1
            sol_numero = f"SOL-{year}-{n_sol:04d}"
            n_oc = (conn.execute(
                "SELECT COALESCE(MAX(CAST(SUBSTR(numero_oc,9) AS INTEGER)),0) FROM ordenes_compra WHERE numero_oc LIKE ?",
                (f"OC-{year}-%",)
            ).fetchone()[0] or 0) + 1
            num_oc = f"OC-{year}-{n_oc:04d}"

            mps_resumen_items = sorted(items, key=lambda x: -x['deficit_g'])[:5]
            mps_resumen = ', '.join([
                (it['nombre'][:25] + f" ({int(it['deficit_g']):,} g)".replace(',', '.'))
                for it in mps_resumen_items
            ])
            if len(items) > 5:
                mps_resumen += f' +{len(items) - 5} más'

            es_sin_prov = (prov == SIN_PROV)
            obs_prefix = (
                "REQUIERE ASIGNAR PROVEEDOR — " if es_sin_prov else
                "Auto-generada Centro Programación — "
            )
            obs = f"{obs_prefix}Proveedor: {prov} · {len(items)} MPs · {mps_resumen}"
            if es_sin_prov:
                obs += " · ACCIÓN: Catalina debe asignar proveedor en /admin → Catálogo MPs y disparar Generar OC nuevamente."

            conn.execute("""INSERT INTO solicitudes_compra
                (numero, fecha, estado, solicitante, urgencia, observaciones,
                 area, empresa, categoria, tipo, numero_oc)
                VALUES (?, datetime('now', '-5 hours'), 'Pendiente', ?, 'Alta', ?,
                        'Produccion', 'Espagiria', 'Materia Prima', 'Compra', ?)""",
                (sol_numero, user, obs, num_oc))

            conn.execute("""INSERT INTO ordenes_compra
                (numero_oc, fecha, estado, proveedor, valor_total, observaciones, creado_por, categoria)
                VALUES (?, datetime('now', '-5 hours'), 'Borrador', ?, 0, ?, ?, 'MP')""",
                (num_oc, prov,
                 f"OC sugerida desde Centro Programación · {len(items)} MPs",
                 user))

            items_created_this = []
            for it in items:
                just = f"Déficit para producción de: {', '.join(it['productos'][:3])}"
                if len(it['productos']) > 3:
                    just += f' +{len(it["productos"])-3} más'
                conn.execute("""INSERT INTO solicitudes_compra_items
                    (numero, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion)
                    VALUES (?,?,?,?,?,?)""",
                    (sol_numero, it['codigo'], it['nombre'],
                     it['deficit_g'], 'g', just))
                try:
                    conn.execute("""INSERT INTO ordenes_compra_items
                        (numero_oc, codigo_mp, nombre_mp, cantidad_g, precio_unitario, subtotal)
                        VALUES (?,?,?,?,0,0)""",
                        (num_oc, it['codigo'], it['nombre'], it['deficit_g']))
                except sqlite3.OperationalError:
                    pass
                items_created_this.append({
                    'codigo': it['codigo'], 'nombre': it['nombre'],
                    'deficit_g': round(it['deficit_g'], 1),
                })

            creadas.append({
                'solicitud': sol_numero, 'orden_compra': num_oc,
                'proveedor': prov, 'n_items': len(items_created_this),
                'requiere_asignacion': es_sin_prov,
            })
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({
            'error': 'Borrado OK pero falló la regeneración',
            'detalle': str(e),
            'borradas': deleted,
        }), 500

    n_total_mps = sum(len(g) for g in grupos.values())
    n_proveedores = len([p for p in grupos.keys() if p != SIN_PROV])
    n_huerfanos = len(grupos.get(SIN_PROV, []))

    return jsonify({
        'ok': True,
        'borradas': deleted,
        'creadas': creadas,
        'total_solicitudes_nuevas': len(creadas),
        'total_mps': n_total_mps,
        'proveedores_resueltos': n_proveedores,
        'mps_huerfanos': n_huerfanos,
        'mensaje': (
            f"✅ Borradas {deleted['solicitudes']} solicitudes viejas + "
            f"creadas {len(creadas)} nuevas con datos actuales "
            f"({n_total_mps} MPs, {n_proveedores} proveedores"
            + (f", {n_huerfanos} huérfanos" if n_huerfanos else "") + ")"
        ),
    })


@bp.route('/api/programacion/generar-oc', methods=['POST'])
def prog_generar_oc():
    """Genera solicitudes de compra agrupadas POR PROVEEDOR.

    Política operacional definida con el dueño del negocio:
      - 1 solicitud por cada proveedor que tenga MPs en déficit
      - 1 solicitud SEPARADA con todas las MPs sin proveedor asignado
        (proveedor='Sin asignar') para que Catalina las revise, asigne en
        /admin → Catálogo MPs, y al refrescar la próxima vez ya queden
        agrupadas en sus proveedores correctos.

    Lee el proveedor de cada MP desde maestro_mps.proveedor.
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401

    conn = get_db()

    # Cálculo correcto: agrega total_g por MP y resta stock UNA vez.
    # Antes se sumaban deficits per-product (under-cuenta cuando hay stock
    # parcial). Esto deja las cantidades del OC alineadas con lo que muestra
    # /programacion (planificacion_estrategica + mps-deficit).
    mp_deficit = _compute_mp_deficit_aggregated(conn, days_ahead=90)

    if not mp_deficit:
        return jsonify({
            'ok': True,
            'mensaje': 'No hay déficits de MP — sin OC necesaria',
            'creadas': []
        })

    # Agrupar MPs en déficit por proveedor; sin proveedor → '(Sin asignar)'
    SIN_PROV = 'Sin asignar'
    grupos = {}  # proveedor → lista de items
    for cod, info in mp_deficit.items():
        prov = info.get('proveedor') or SIN_PROV
        grupos.setdefault(prov, []).append({
            'codigo': cod,
            'nombre': info['nombre'],
            'deficit_g': info['deficit_g'],
            'productos': info['productos'],
        })

    # Numero de SOL inicial
    from datetime import datetime as _dt
    year = _dt.now().strftime('%Y')
    last_n = conn.execute(
        "SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) FROM solicitudes_compra WHERE numero LIKE ?",
        (f"SOL-{year}-%",)
    ).fetchone()[0] or 0
    n_sol = last_n
    user = session.get('compras_user', 'Sistema')

    creadas = []
    huerfana_info = None

    # Ordenar: primero proveedores reales, al final 'Sin asignar' para que
    # quede visualmente claro como solicitud "por revisar"
    proveedores_ordenados = sorted(
        grupos.keys(),
        key=lambda p: (p == SIN_PROV, p.lower())
    )

    try:
        for prov in proveedores_ordenados:
            items = grupos[prov]
            if not items:
                continue
            n_sol += 1
            sol_numero = f"SOL-{year}-{n_sol:04d}"
            n_oc = (conn.execute(
                "SELECT COALESCE(MAX(CAST(SUBSTR(numero_oc,9) AS INTEGER)),0) FROM ordenes_compra WHERE numero_oc LIKE ?",
                (f"OC-{year}-%",)
            ).fetchone()[0] or 0) + 1
            num_oc = f"OC-{year}-{n_oc:04d}"

            # Resumen de MPs principales para observaciones (legibilidad)
            mps_resumen_items = sorted(items, key=lambda x: -x['deficit_g'])[:5]
            mps_resumen = ', '.join([
                (it['nombre'][:25] + f" ({int(it['deficit_g']):,} g)".replace(',', '.'))
                for it in mps_resumen_items
            ])
            if len(items) > 5:
                mps_resumen += f' +{len(items) - 5} más'

            es_sin_prov = (prov == SIN_PROV)
            obs_prefix = (
                "REQUIERE ASIGNAR PROVEEDOR — " if es_sin_prov else
                "Auto-generada Centro Programación — "
            )
            obs = (
                f"{obs_prefix}Proveedor: {prov} · "
                f"{len(items)} MPs · {mps_resumen}"
            )
            if es_sin_prov:
                obs += " · ACCIÓN: Catalina debe asignar proveedor en /admin → Catálogo MPs y disparar Generar OC nuevamente."

            conn.execute("""INSERT INTO solicitudes_compra
                (numero, fecha, estado, solicitante, urgencia, observaciones,
                 area, empresa, categoria, tipo, numero_oc)
                VALUES (?, datetime('now', '-5 hours'), 'Pendiente', ?, 'Alta', ?,
                        'Produccion', 'Espagiria', 'Materia Prima', 'Compra', ?)""",
                (sol_numero, user, obs, num_oc))

            # OC asociada con el proveedor correcto (o 'Sin asignar')
            valor_estimado_total = 0
            conn.execute("""INSERT INTO ordenes_compra
                (numero_oc, fecha, estado, proveedor, valor_total, observaciones, creado_por, categoria)
                VALUES (?, datetime('now', '-5 hours'), 'Borrador', ?, 0, ?, ?, 'MP')""",
                (num_oc, prov,
                 f"OC sugerida desde Centro Programación · {len(items)} MPs",
                 user))

            items_created_this = []
            for it in items:
                just = f"Déficit para producción de: {', '.join(it['productos'][:3])}"
                if len(it['productos']) > 3:
                    just += f' +{len(it["productos"])-3} más'
                conn.execute("""INSERT INTO solicitudes_compra_items
                    (numero, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion)
                    VALUES (?,?,?,?,?,?)""",
                    (sol_numero, it['codigo'], it['nombre'],
                     it['deficit_g'], 'g', just))
                try:
                    conn.execute("""INSERT INTO ordenes_compra_items
                        (numero_oc, codigo_mp, nombre_mp, cantidad_g, precio_unitario, subtotal)
                        VALUES (?,?,?,?,0,0)""",
                        (num_oc, it['codigo'], it['nombre'], it['deficit_g']))
                except sqlite3.OperationalError:
                    pass
                items_created_this.append({
                    'codigo': it['codigo'],
                    'nombre': it['nombre'],
                    'deficit_g': round(it['deficit_g'], 1),
                })

            entry = {
                'solicitud': sol_numero,
                'orden_compra': num_oc,
                'proveedor': prov,
                'n_items': len(items_created_this),
                'items': items_created_this,
                'requiere_asignacion': es_sin_prov,
            }
            creadas.append(entry)
            if es_sin_prov:
                huerfana_info = entry

        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({
            'error': 'Falló la generación de OCs',
            'detalle': str(e),
        }), 500

    n_total_mps = sum(len(g) for g in grupos.values())
    n_proveedores = len([p for p in grupos.keys() if p != SIN_PROV])
    n_huerfanos = len(grupos.get(SIN_PROV, []))

    msg = f'{len(creadas)} solicitudes creadas — {n_total_mps} MPs en {n_proveedores} proveedores'
    if n_huerfanos:
        msg += f' + 1 solicitud SIN ASIGNAR ({n_huerfanos} MPs huérfanos para revisar)'

    return jsonify({
        'ok': True,
        'creadas': creadas,
        'total_solicitudes': len(creadas),
        'total_mps': n_total_mps,
        'proveedores_resueltos': n_proveedores,
        'mps_huerfanos': n_huerfanos,
        'huerfana': huerfana_info,
        'mensaje': msg,
    })


@bp.route('/api/programacion/mps-deficit')
def prog_mps_deficit():
    """Lista plana de MPs con déficit real según Centro de Programación.

    A diferencia de /api/alertas-reabastecimiento (que compara stock_actual
    < stock_minimo y depende de un campo desactualizado), esta calcula
    déficit REAL: necesidad para producciones futuras + velocidad de ventas
    proyectada vs stock actual.

    Cada item: codigo_mp, nombre, stock_actual_g, deficit_g, productos_afectados,
                proveedor, lead_time_estimado.
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db()
    try:
        china_mps_set = _get_china_mps(conn)
        # Misma fuente de verdad que /generar-oc y /planificacion
        mp_def_raw = _compute_mp_deficit_aggregated(conn, days_ahead=90)
        if not mp_def_raw:
            return jsonify({'mps': [], 'total': 0, 'deficit_total_kg': 0})

        items = []
        for mid, info in mp_def_raw.items():
            items.append({
                'codigo_mp': mid,
                'nombre': info['nombre'],
                'stock_actual_g': info['stock_g'],
                'deficit_g': info['deficit_g'],
                'productos_afectados': info['productos'],
                'es_china': mid in china_mps_set,
                'proveedor': info.get('proveedor', ''),
            })
        items.sort(key=lambda x: -x['deficit_g'])
        deficit_total_kg = sum(i['deficit_g'] for i in items) / 1000.0

        return jsonify({
            'mps': items,
            'total': len(items),
            'deficit_total_kg': round(deficit_total_kg, 2),
            'china_count': sum(1 for i in items if i.get('es_china')),
        })
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()[-500:],
            'mps': [],
            'total': 0,
        }), 500


@bp.route('/api/programacion/n-alertas')
def prog_n_alertas():
    """Endpoint rápido — retorna solo el conteo de alertas (para badge en Compras)."""
    if not _auth():
        return jsonify({'n': 0})
    conn = get_db()
    try:
        vel_data = _shopify_velocity(conn, days=60)
        mp_stock = _get_mp_stock(conn)
        formulas = _get_formulas(conn)
        if not formulas:
            return jsonify({'n': 0, 'criticos': 0})
        _, alerts = _project_stock(
            conn, vel_data['prod_velocity'], formulas, mp_stock, []
        )
        criticos = len([a for a in alerts if a['nivel'] == 'critico'])
        return jsonify({'n': len(alerts), 'criticos': criticos})
    except Exception as e:
        return jsonify({'n': 0, 'error': str(e)})


@bp.route('/api/programacion/debug-mps')
def prog_debug_mps():
    """Debug: cross-reference mp_stock entries vs formula_items material_ids."""
    if not _auth():
        return jsonify({'error': 'no auth'}), 401
    conn = get_db()

    # MP stock from movimientos
    mp_stock = _get_mp_stock(conn)

    # All material_ids used in formulas
    formula_mids = conn.execute(
        "SELECT DISTINCT material_id, material_nombre FROM formula_items ORDER BY material_id"
    ).fetchall()

    matched   = []
    unmatched = []
    for mid, nombre in formula_mids:  # nombre = material_nombre
        # Try: unlimited → ID → exact name → norm name → alias
        nombre_str = str(nombre or '')
        if _is_unlimited_mp(nombre_str):
            stock = float('inf')
            match_via = 'unlimited'
        elif mid in mp_stock:
            stock = mp_stock[mid]
            match_via = 'id'
        else:
            nombre_exact = nombre_str.strip().upper()
            if nombre_exact in mp_stock:
                stock = mp_stock[nombre_exact]
                match_via = 'nombre_exact'
            else:
                nombre_norm = _norm_mp_name(nombre_str)
                if nombre_norm in mp_stock:
                    stock = mp_stock[nombre_norm]
                    match_via = 'nombre_norm'
                else:
                    alias_key = _MP_NAME_ALIAS.get(nombre_norm) or _MP_NAME_ALIAS.get(nombre_exact)
                    if alias_key and alias_key in mp_stock:
                        stock = mp_stock[alias_key]
                        match_via = 'alias'
                    else:
                        stock = None
                        match_via = None

        if stock is not None:
            matched.append({'material_id': mid, 'nombre': nombre,
                            'stock_g': round(stock, 1), 'match_via': match_via})
        else:
            unmatched.append({'material_id': mid, 'nombre': nombre})

    # Sample mp_stock keys (first 20)
    sample_stock_keys = sorted(mp_stock.keys())[:20]

    # Movimientos table row count
    n_mov = conn.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]

    # Check if movimientos has any 'entrada' tipo
    n_entradas = conn.execute(
        "SELECT COUNT(*) FROM movimientos WHERE tipo IN ('entrada','Entrada','ENTRADA')"
    ).fetchone()[0]

    # Search mp_stock for critical keywords
    keywords = ['AGUA', 'ALOE', 'ARGAN', 'JOJOBA', 'ASCORBICO', 'FERULICO',
                'PANTENOL', 'CETILICO', 'CETEARIL', 'ROSA MOSQUETA',
                'ACETIL', 'ACETYL', 'HIALURONICO',
                'SILICONA', 'EZ', 'REGALIZ', 'GLYCYRRHIZA', 'LICORICE',
                'CENTELLA', 'PROPILEN', 'CARBOPOL', 'CAFEINA', 'CERAMIDA',
                'BETAGLUCAN', 'BACKUCHIOL', 'NIACINAMIDA', 'RETINOL']
    keyword_hits = {}
    for kw in keywords:
        hits = [k for k in mp_stock if kw in k]
        if hits:
            keyword_hits[kw] = hits[:5]

    # Show normalized form of unmatched names for debugging
    unmatched_with_norm = []
    for item in unmatched:
        norm_key = _norm_mp_name(item['nombre'])
        in_stock = norm_key in mp_stock
        unmatched_with_norm.append({**item, 'norm_key': norm_key, 'norm_in_stock': in_stock})

    return jsonify({
        'mp_stock_total_entries': len(mp_stock),
        'mp_stock_sample_keys': sample_stock_keys,
        'formula_items_distinct_materials': len(formula_mids),
        'matched_count': len(matched),
        'unmatched_count': len(unmatched),
        'unmatched': unmatched_with_norm[:30],
        'matched_sample': matched[:10],
        'movimientos_total_rows': n_mov,
        'movimientos_entradas': n_entradas,
        'keyword_hits_in_stock': keyword_hits,
    })


@bp.route('/api/programacion/debug-mp-check/<producto>')
def prog_debug_mp_check(producto):
    """Debug: muestra el mp_check completo para un producto específico."""
    if not _auth():
        return jsonify({'error': 'no auth'}), 401
    conn = get_db()
    mp_stock = _get_mp_stock(conn)
    formulas = _get_formulas(conn)

    prod_key = None
    for k in formulas:
        if k.upper() == producto.upper() or producto.upper() in k.upper():
            prod_key = k
            break

    if not prod_key:
        available = list(formulas.keys())
        return jsonify({'error': f'Producto no encontrado', 'disponibles': available[:20]}), 404

    data = formulas[prod_key]
    lote_kg = data.get('lote_size_kg') or 1.0
    items = data.get('items', [])
    items_with_qty = [i for i in items if i.get('cantidad_g_por_lote', 0) > 0]

    result = []
    for item in items_with_qty:
        mid = str(item['material_id']).strip()
        nombre_raw = str(item.get('material_nombre', '')).strip()
        needed_g = float(item['cantidad_g_por_lote'])

        # Mirror exact lookup logic from _project_stock
        if _is_unlimited_mp(nombre_raw):
            available_g = float('inf')
            match_via = 'unlimited'
        elif mid in mp_stock:
            available_g = mp_stock[mid]
            match_via = 'id'
        else:
            nombre_exact = nombre_raw.upper()
            if nombre_exact in mp_stock:
                available_g = mp_stock[nombre_exact]
                match_via = 'nombre_exact'
            else:
                nombre_norm = _norm_mp_name(nombre_raw)
                if nombre_norm in mp_stock:
                    available_g = mp_stock[nombre_norm]
                    match_via = 'nombre_norm'
                else:
                    alias_key = _MP_NAME_ALIAS.get(nombre_norm) or _MP_NAME_ALIAS.get(nombre_exact)
                    if alias_key and alias_key in mp_stock:
                        available_g = mp_stock[alias_key]
                        match_via = f'alias→{alias_key}'
                    else:
                        available_g = 0
                        match_via = 'NOT_FOUND'

        deficit_g = 0 if available_g == float('inf') else max(0, needed_g - available_g)
        ok = deficit_g < 1
        result.append({
            'material_id': mid,
            'nombre': nombre_raw,
            'needed_g': round(needed_g, 1),
            'available_g': '∞' if available_g == float('inf') else round(available_g, 1),
            'deficit_g': round(deficit_g, 1),
            'ok': ok,
            'match_via': match_via,
        })

    result.sort(key=lambda x: (x['ok'], -x['deficit_g']))  # failing first
    failing = [r for r in result if not r['ok']]
    return jsonify({
        'producto': prod_key,
        'lote_kg': lote_kg,
        'total_ingredientes': len(result),
        'ingredientes_faltantes': len(failing),
        'can_produce': len(failing) == 0,
        'failing_first': result,
    })


# ─── MP Bridge — admin endpoints ─────────────────────────────────────────────

@bp.route('/api/programacion/mp-bridge', methods=['GET'])
def mp_bridge_list():
    """List all bridge mappings (active and inactive)."""
    with _db() as conn:
        rows = conn.execute("""
            SELECT id, formula_material_id, formula_material_nombre,
                   bodega_material_id, bodega_material_nombre, bodega_inci,
                   notas, activo, creado_en
            FROM mp_formula_bridge
            ORDER BY formula_material_nombre NULLS LAST
        """).fetchall()
        keys = ['id', 'formula_material_id', 'formula_material_nombre',
                'bodega_material_id', 'bodega_material_nombre', 'bodega_inci',
                'notas', 'activo', 'creado_en']
        return jsonify([dict(zip(keys, r)) for r in rows])


@bp.route('/api/programacion/mp-bridge', methods=['POST'])
def mp_bridge_add():
    """Add or update a bridge mapping.

    Body JSON: {
        formula_material_id, formula_material_nombre,
        bodega_material_id, bodega_material_nombre, bodega_inci, notas
    }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json(force=True) or {}
    fid   = str(data.get('formula_material_id', '') or '').strip()
    fname = str(data.get('formula_material_nombre', '') or '').strip()
    bid   = str(data.get('bodega_material_id', '') or '').strip()
    bname = str(data.get('bodega_material_nombre', '') or '').strip()
    binci = str(data.get('bodega_inci', '') or '').strip()
    notas = str(data.get('notas', '') or '').strip()

    if not fid or not bid:
        return jsonify({'error': 'formula_material_id y bodega_material_id son obligatorios'}), 400

    with _db() as conn:
        conn.execute("""
            INSERT INTO mp_formula_bridge
                (formula_material_id, formula_material_nombre,
                 bodega_material_id, bodega_material_nombre, bodega_inci, notas, activo)
            VALUES (?, ?, ?, ?, ?, ?, 1)
            ON CONFLICT(formula_material_id, bodega_material_id)
            DO UPDATE SET
                formula_material_nombre = excluded.formula_material_nombre,
                bodega_material_nombre  = excluded.bodega_material_nombre,
                bodega_inci             = excluded.bodega_inci,
                notas                   = excluded.notas,
                activo                  = 1
        """, (fid, fname or None, bid, bname or None, binci or None, notas or None))
        conn.commit()
    return jsonify({'ok': True, 'formula_material_id': fid, 'bodega_material_id': bid})


@bp.route('/api/programacion/mp-bridge/<int:bridge_id>', methods=['DELETE'])
def mp_bridge_delete(bridge_id):
    """Soft-delete a bridge mapping (sets activo=0)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    with _db() as conn:
        conn.execute("UPDATE mp_formula_bridge SET activo=0 WHERE id=?", (bridge_id,))
        conn.commit()
    return jsonify({'ok': True, 'id': bridge_id, 'activo': 0})


@bp.route('/api/programacion/mp-bridge/unmatched', methods=['GET'])
def mp_bridge_unmatched():
    """Return formula_items that cannot be matched to any movimientos entry.

    Useful for populating the bridge table: shows what still needs linking,
    along with candidate bodega materials (fuzzy name match by keywords).
    """
    with _db() as conn:
        mp_stock = _get_mp_stock(conn)

        # All bridge mappings already in place (formula_material_id → bodega_material_id)
        bridged_fids = set(r[0] for r in conn.execute(
            "SELECT formula_material_id FROM mp_formula_bridge WHERE activo=1"
        ).fetchall())

        # Fetch distinct formula items (one entry per material_id)
        items = conn.execute("""
            SELECT DISTINCT material_id, material_nombre
            FROM formula_items
            WHERE material_id IS NOT NULL AND material_id != ''
            ORDER BY material_nombre
        """).fetchall()

        # Fetch all bodega materials for candidate suggestions
        bodega_mats = conn.execute("""
            SELECT DISTINCT material_id, material_nombre
            FROM movimientos
            WHERE material_id IS NOT NULL AND material_id != ''
              AND material_nombre IS NOT NULL AND material_nombre != ''
            ORDER BY material_nombre
        """).fetchall()

        unmatched = []
        for mid, nombre in items:
            mid = str(mid or '').strip()
            nombre_raw = str(nombre or '').strip()

            # Skip already bridged
            if mid in bridged_fids:
                continue
            # Skip unlimited MPs
            if _is_unlimited_mp(nombre_raw):
                continue
            # Check if already matched via stock dict
            nombre_exact = nombre_raw.upper()
            nombre_norm  = _norm_mp_name(nombre_raw)
            alias_key    = _MP_NAME_ALIAS.get(nombre_norm) or _MP_NAME_ALIAS.get(nombre_exact)

            already_matched = (
                mid in mp_stock
                or nombre_exact in mp_stock
                or nombre_norm in mp_stock
                or (alias_key and alias_key in mp_stock)
            )
            if already_matched:
                continue

            # Build candidate suggestions: bodega materials whose normalized name
            # shares at least one 4-char keyword with the formula material name
            formula_keywords = set(w for w in nombre_norm.split() if len(w) >= 4)
            candidates = []
            for bid, bname in bodega_mats:
                bname_norm = _norm_mp_name(str(bname or ''))
                bname_words = set(w for w in bname_norm.split() if len(w) >= 4)
                shared = formula_keywords & bname_words
                if shared:
                    candidates.append({
                        'material_id': bid,
                        'material_nombre': bname,
                        'shared_keywords': sorted(shared),
                        'score': len(shared)
                    })
            candidates.sort(key=lambda x: -x['score'])

            unmatched.append({
                'formula_material_id':     mid,
                'formula_material_nombre': nombre_raw,
                'candidates':              candidates[:5]  # top 5 suggestions
            })

        return jsonify({
            'total_unmatched': len(unmatched),
            'unmatched': unmatched
        })


# ─── Producciones con faltantes (vista simple para Luis Enrique) ─────────
# Sebastian 5-may-2026: jefe de producción quiere UN solo flujo.
#   1. Selector horizonte (semana, 15d, 1m, 2m, 3m, 6m, 1 año)
#   2. Tabla de producciones programadas en ese horizonte
#   3. Por cada producción: ver MP+MEE faltantes
#   4. Botón "Solicitar TODO faltante" → bulk OC por proveedor
#
# Reutiliza:
#   - produccion_programada (eventos programados)
#   - formula_items (MP por producto)
#   - sku_mee_config (MEE por producto)
#   - maestro_mps + mp_lead_time_config (proveedor sugerido MP)
#   - maestro_mee + mee_lead_time_config (proveedor sugerido MEE)
#
# Excluye producciones con inventario_descontado_at != NULL (ya consumieron).

@bp.route('/api/programacion/producciones-faltantes', methods=['GET'])
def producciones_faltantes():
    """Producciones programadas en horizonte + agregado de MP/MEE faltantes.

    Query params:
      dias: int 7-365 (default 60)

    Returns:
      {
        horizonte_dias: int,
        producciones: [
          {id, producto, fecha, lotes, cantidad_kg,
           mps_necesarias: [{codigo_mp, nombre, necesario_g}],
           mees_necesarios: [{codigo, descripcion, tipo, necesario_unidades}]
          }
        ],
        faltantes_mps: [
          {codigo_mp, nombre, necesario_total_g, stock_actual_g, faltante_g,
           proveedor_sugerido}
        ],
        faltantes_mees: [
          {codigo, descripcion, tipo, necesario_total_u, stock_actual_u,
           faltante_u, proveedor_sugerido}
        ],
        resumen: {n_producciones, n_mps_faltantes, n_mees_faltantes,
                  n_proveedores_unicos}
      }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = max(7, min(int(request.args.get('dias', 60)), 365))
    except (ValueError, TypeError):
        dias = 60

    # Sebastián 9-may-2026: las producciones programadas que pasaron pero
    # nunca arrancaron (sin inicio_real_at, sin descontar) ensucian el panel
    # con basura vieja ("Mar 28 ya pasó", "Lun 4 ya pasó", ...) que ya no
    # se va a producir. Antes del fix se mezclaba con lo real.
    #
    # Ahora se filtra:
    #   - REALIZADAS pasadas (con inicio_real_at o desc_at): SIEMPRE se
    #     muestran independiente del horizonte → trazabilidad histórica.
    #   - PENDIENTES pasadas: solo se muestran si fecha >= hoy - atrasadas_max_dias
    #     (default 7d). Una atrasada de >7d sin arrancar es ruido.
    #   - Si el usuario quiere ver TODAS las atrasadas pendientes (para
    #     limpiarlas/cancelarlas), pasa ?atrasadas_max_dias=999 desde el
    #     toggle "Mostrar atrasadas viejas" del frontend.
    try:
        atrasadas_max_dias = max(0, min(int(request.args.get('atrasadas_max_dias', 7)), 999))
    except (ValueError, TypeError):
        atrasadas_max_dias = 7

    conn = get_db(); c = conn.cursor()
    from datetime import date, timedelta as _td
    hoy = date.today()
    cutoff = (hoy + _td(days=dias)).isoformat()
    # Ventana pasada amplia (14d) solo para que las realizadas viejas SÍ aparezcan.
    # Las pendientes-pasadas se filtran por separado abajo.
    pasado_dias = 14
    past_window = (hoy - _td(days=pasado_dias)).isoformat()
    # Threshold para pendientes-pasadas (atrasadas que no arrancaron)
    atrasadas_cutoff = (hoy - _td(days=atrasadas_max_dias)).isoformat()

    # 1. Cargar producciones programadas pendientes Y realizadas en ventana.
    # Sebastián 9-may-2026: condición compuesta para evitar ruido de atrasadas
    # viejas pendientes. La pasada-pendiente (sin inicio_real_at, sin desc_at)
    # solo aparece si fecha >= atrasadas_cutoff (default hoy-7d).
    # La REALIZADA (tiene inicio_real_at o desc_at) siempre aparece dentro de
    # la ventana past_window (14d hacia atrás · trazabilidad reciente).
    # La FUTURA (>= hoy) siempre aparece hasta el cutoff.
    try:
        c.execute("""
            SELECT pp.id, pp.producto, pp.fecha_programada,
                   COALESCE(pp.lotes, 1), COALESCE(pp.cantidad_kg, 0),
                   COALESCE(fh.lote_size_kg, 0) as lote_size_kg,
                   COALESCE(pp.area_id, 0) as area_id,
                   COALESCE(ap.codigo, '') as area_codigo,
                   COALESCE(ap.nombre, '') as area_nombre,
                   COALESCE(LOWER(pp.estado), 'pendiente') as estado_norm,
                   COALESCE(pp.inicio_real_at, '') as inicio_real_at,
                   COALESCE(pp.fin_real_at, '') as fin_real_at,
                   COALESCE(pp.inventario_descontado_at, '') as desc_at
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh
                   ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            LEFT JOIN areas_planta ap ON ap.id = pp.area_id
            WHERE LOWER(COALESCE(pp.estado, '')) != 'cancelado'
              AND pp.fecha_programada >= ?
              AND pp.fecha_programada <= ?
              AND (
                    -- Realizada: con inicio_real_at o desc_at → siempre incluir
                    COALESCE(pp.inicio_real_at, '') != ''
                 OR COALESCE(pp.inventario_descontado_at, '') != ''
                    -- Futura o atrasada reciente (≤ atrasadas_max_dias)
                 OR pp.fecha_programada >= ?
              )
            ORDER BY pp.fecha_programada ASC
        """, (past_window, cutoff, atrasadas_cutoff))
        prod_rows = c.fetchall()
    except sqlite3.OperationalError:
        # Fallback si areas_planta no existe (esquema legacy)
        c.execute("""
            SELECT pp.id, pp.producto, pp.fecha_programada,
                   COALESCE(pp.lotes, 1), COALESCE(pp.cantidad_kg, 0),
                   COALESCE(fh.lote_size_kg, 0) as lote_size_kg,
                   0, '', '',
                   COALESCE(LOWER(pp.estado), 'pendiente'),
                   COALESCE(pp.inicio_real_at, ''),
                   COALESCE(pp.fin_real_at, ''),
                   COALESCE(pp.inventario_descontado_at, '')
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh
                   ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE LOWER(COALESCE(pp.estado, '')) != 'cancelado'
              AND pp.fecha_programada >= ?
              AND pp.fecha_programada <= ?
              AND (
                    COALESCE(pp.inicio_real_at, '') != ''
                 OR COALESCE(pp.inventario_descontado_at, '') != ''
                 OR pp.fecha_programada >= ?
              )
            ORDER BY pp.fecha_programada ASC
        """, (past_window, cutoff, atrasadas_cutoff))
        prod_rows = c.fetchall()

    # 2. Cargar formulas (codigo_mp -> [nombre, cantidad_g_por_lote, porcentaje])
    formulas = {}
    for r in c.execute("""
        SELECT producto_nombre, material_id, material_nombre,
               COALESCE(porcentaje, 0), COALESCE(cantidad_g_por_lote, 0)
        FROM formula_items
    """).fetchall():
        prod = (r[0] or '').strip().upper()
        formulas.setdefault(prod, []).append({
            'codigo_mp': r[1] or '',
            'nombre': r[2] or '',
            'porcentaje': float(r[3] or 0),
            'g_por_lote': float(r[4] or 0),
        })

    # 3. Cargar sku_mee_config (producto/sku -> [mee_codigo, cantidad_por_unidad, tipo])
    mee_por_producto = {}
    try:
        for r in c.execute("""
            SELECT s.sku_codigo, s.mee_codigo,
                   COALESCE(s.cantidad_por_unidad, 1),
                   COALESCE(s.componente_tipo, 'envase'),
                   COALESCE(m.descripcion, '') as descripcion
            FROM sku_mee_config s
            LEFT JOIN maestro_mee m ON m.codigo = s.mee_codigo
            WHERE COALESCE(s.aplica, 1) = 1
        """).fetchall():
            sku = (r[0] or '').strip().upper()
            mee_por_producto.setdefault(sku, []).append({
                'mee_codigo': r[1] or '',
                'cantidad_por_unidad': float(r[2] or 1),
                'tipo': r[3] or 'envase',
                'descripcion': r[4] or '',
            })
    except sqlite3.OperationalError:
        # Tabla no existe (esquema antiguo) · seguir sin MEE
        mee_por_producto = {}

    # 4. Cargar volumen unitario por producto (ml por unidad envasada)
    volumen_por_producto = {}
    try:
        for r in c.execute("""
            SELECT producto_nombre, COALESCE(volumen_ml, 0)
            FROM volumen_unitario_producto WHERE COALESCE(activo,1)=1
        """).fetchall():
            volumen_por_producto[(r[0] or '').strip().upper()] = float(r[1] or 0)
    except sqlite3.OperationalError:
        pass
    # Si no hay tabla volumen_unitario, intentar producto_presentaciones
    if not volumen_por_producto:
        try:
            for r in c.execute("""
                SELECT producto_nombre, MAX(COALESCE(volumen_ml, 0))
                FROM producto_presentaciones
                WHERE COALESCE(activo,1)=1
                GROUP BY producto_nombre
            """).fetchall():
                volumen_por_producto[(r[0] or '').strip().upper()] = float(r[1] or 0)
        except sqlite3.OperationalError:
            pass

    # FASE 2 · 27-may-2026 PM · cargar TODAS las presentaciones por producto
    # cuando hay >1 (caso variantes 30ml/15ml/10ml) · cálculo correcto sumando
    # envases por presentación con ratio histórico Shopify 90d.
    # Estructura: presentaciones_por_producto[producto_norm] = [
    #   {'codigo','volumen_ml','envase_codigo','sku_shopify','factor_g_por_unidad'}, ...
    # ]
    presentaciones_por_producto = {}
    try:
        for r in c.execute("""
            SELECT producto_nombre, presentacion_codigo,
                   COALESCE(volumen_ml,0), COALESCE(envase_codigo,''),
                   COALESCE(sku_shopify,''),
                   COALESCE(factor_g_por_unidad, volumen_ml, 0),
                   COALESCE(ventas_mes_referencia, 0)
            FROM producto_presentaciones
            WHERE COALESCE(activo,1)=1 AND COALESCE(volumen_ml,0) > 0
        """).fetchall():
            pn = (r[0] or '').strip().upper()
            if not pn:
                continue
            presentaciones_por_producto.setdefault(pn, []).append({
                'codigo': r[1],
                'volumen_ml': float(r[2] or 0),
                'envase_codigo': (r[3] or '').strip(),
                'sku_shopify': (r[4] or '').strip(),
                'factor_g_por_unidad': float(r[5] or r[2] or 0),
                'ventas_mes_referencia': float(r[6] or 0),
            })
    except sqlite3.OperationalError:
        pass

    # Ratio histórico Shopify por presentación (últimos 180d · 6m) Sebastián
    # 27-may-2026 PM · "si lanzas la busqueda de ventas en 6 meses sabras que
    # se vendio no puede se 50 50". Ampliado de 90d para capturar SKUs B2B
    # con cadencia mensual o trimestral.
    # Si Shopify sync no está activo o no hay data → ratio uniforme (1/N).
    ventas_por_sku_90d = {}
    try:
        from datetime import datetime as _dt90, timedelta as _td90
        _cutoff_90d = (_dt90.utcnow() - _td90(hours=5) - _td90(days=180)).date().isoformat()
        import json as _json_sk
        rows_sho = c.execute(
            """SELECT sku_items FROM animus_shopify_orders
               WHERE COALESCE(creado_en,'') >= ? AND sku_items IS NOT NULL AND sku_items != ''""",
            (_cutoff_90d,)
        ).fetchall()
        for (it,) in rows_sho:
            try:
                items = _json_sk.loads(it) if it else []
                if not isinstance(items, list):
                    continue
                for li in items:
                    sk = (li.get('sku') or '').strip()
                    qty = int(li.get('qty') or 0)
                    if sk and qty > 0:
                        ventas_por_sku_90d[sk] = ventas_por_sku_90d.get(sk, 0) + qty
            except Exception:
                continue
    except sqlite3.OperationalError:
        pass

    def _ratio_presentaciones(producto_norm):
        """Devuelve dict {codigo_presentacion: ratio_0_a_1} con esta prioridad:
        1) Override manual: si AT LEAST 1 presentación tiene ventas_mes_referencia > 0,
           usa ESOS números (0 = no se vende).
        2) Shopify histórico: ratio = qty_sku / total_qty últimos 180d.
        3) Uniforme: 1/N si no hay data.
        Sebastián 27-may-2026 PM · "AZ lo vendemos de 30 y 15, de 15 200 uds/mes"."""
        pres = presentaciones_por_producto.get(producto_norm, [])
        if not pres:
            return {}
        if len(pres) == 1:
            return {pres[0]['codigo']: 1.0}
        # Prioridad 1 · override manual ventas_mes_referencia
        suma_override = sum(float(p.get('ventas_mes_referencia') or 0) for p in pres)
        if suma_override > 0:
            return {p['codigo']: (float(p.get('ventas_mes_referencia') or 0) / suma_override)
                    for p in pres}
        # Prioridad 2 · Shopify
        ventas = {}
        total = 0
        for p in pres:
            sk = (p['sku_shopify'] or '').strip()
            v = ventas_por_sku_90d.get(sk, 0) if sk else 0
            ventas[p['codigo']] = v
            total += v
        if total > 0:
            return {cod: (v / total) for cod, v in ventas.items()}
        # Prioridad 3 · uniforme
        return {p['codigo']: 1.0 / len(pres) for p in pres}

    # 5. Cargar stock MP (canonical helper)
    stock_mp = _get_mp_stock(conn)

    # 6. Cargar stock MEE
    # P0-4 23-may-PM · auditoría agente Stock · antes leía maestro_mee
    # .stock_actual directo (cache que puede drift) · ahora usa el helper
    # canónico _get_mee_stock que suma movimientos_mee.
    stock_mee = _get_mee_stock(conn)

    # 7. Cargar info MP (nombre canonico + proveedor sugerido)
    mp_info = {}
    try:
        for r in c.execute("""
            SELECT mm.codigo_mp,
                   COALESCE(mm.nombre_comercial, mm.nombre_inci, mm.codigo_mp) as nombre,
                   COALESCE(NULLIF(TRIM(mlt.proveedor_principal),''), mm.proveedor, '') as prov
            FROM maestro_mps mm
            LEFT JOIN mp_lead_time_config mlt ON mlt.material_id = mm.codigo_mp
            WHERE COALESCE(mm.activo, 1) = 1
        """).fetchall():
            mp_info[r[0]] = {'nombre': r[1] or r[0], 'proveedor': (r[2] or '').strip()}
    except sqlite3.OperationalError:
        pass

    # 8. Cargar info MEE (proveedor sugerido)
    mee_info = {}
    try:
        for r in c.execute("""
            SELECT mm.codigo,
                   COALESCE(mm.descripcion, mm.codigo) as descripcion,
                   COALESCE(cfg.proveedor_principal, mm.proveedor, '') as prov
            FROM maestro_mee mm
            LEFT JOIN mee_lead_time_config cfg ON cfg.mee_codigo = mm.codigo
        """).fetchall():
            mee_info[r[0]] = {
                'descripcion': r[1] or r[0],
                'proveedor': (r[2] or '').strip(),
            }
    except sqlite3.OperationalError:
        pass

    # 9. Por cada producción, calcular MP y MEE necesarios
    producciones_out = []
    consumo_mp_agregado = {}   # codigo_mp -> g_total
    consumo_mee_agregado = {}  # mee_codigo -> unidades_total
    _claves_vistas = set()     # anti-clon · ver _clave_clon abajo
    productos_sin_volumen = set()  # con MEE configurado pero sin volumen_ml
    hoy_iso = hoy.isoformat()
    for row in prod_rows:
        # Sebastian 5-may-2026: unpack soporta ambos esquemas (con/sin areas_planta)
        # Sebastian 8-may-2026: nuevo campo `desc_at` (inventario_descontado_at)
        # para detectar realizadas.
        desc_at = ''
        if len(row) >= 13:
            (pid, producto, fecha, lotes, cant_kg_explicita, lote_size,
             area_id, area_codigo, area_nombre,
             estado_norm, inicio_real_at, fin_real_at, desc_at) = row
        elif len(row) >= 12:
            (pid, producto, fecha, lotes, cant_kg_explicita, lote_size,
             area_id, area_codigo, area_nombre,
             estado_norm, inicio_real_at, fin_real_at) = row
        else:
            pid, producto, fecha, lotes, cant_kg_explicita, lote_size = row[:6]
            area_id, area_codigo, area_nombre = 0, '', ''
            estado_norm, inicio_real_at, fin_real_at = 'pendiente', '', ''
        producto_norm = (producto or '').strip().upper()
        lotes = int(lotes or 1)
        cant_kg_total = float(cant_kg_explicita or 0) or (lotes * float(lote_size or 0))

        # Sebastián 8-may-2026: clasificación visual + lógica MP separadas.
        # · Realizada: TERMINADA (fin_real_at o estado=completado)
        # · En proceso: arrancó pero no terminó (inicio_real_at sin fin · o
        #   descontada sin fin)
        # · Atrasada: fecha pasó sin arrancar y sin descontar · sigue pendiente
        # · Pendiente: futura sin arrancar · caso normal
        # MPs/MEEs faltantes: regla por separado · NO contar si ya hay descuento
        # registrado (desc_at set) · esos MPs ya salieron del stock.
        es_realizada = bool(fin_real_at) or estado_norm == 'completado'
        es_en_proceso = (bool(inicio_real_at) or bool(desc_at)) and not es_realizada
        es_pasada = (fecha or '') < hoy_iso
        es_atrasada = es_pasada and not es_realizada and not es_en_proceso
        # Dedup anti-clon · si el sync de Calendar insertó un duplicado
        # EXACTO (mismo producto, fecha, lotes y kg), solo el primero cuenta
        # para el agregado de faltantes · si no, se pediría comprar el doble.
        # Clave EXACTA (no la heurística de ventana de 7d, que daría falsos
        # positivos): el scheduling real no repite lo idéntico el mismo día.
        _clave_clon = (producto_norm, (fecha or '')[:10], lotes, round(cant_kg_total, 3))
        es_clon = _clave_clon in _claves_vistas
        if not es_clon:
            _claves_vistas.add(_clave_clon)
        # Solo sumar al consumo si NO se ha descontado, aún hay que hacerla
        # y NO es un clon exacto ya contado.
        cuenta_para_faltantes = not bool(desc_at) and not es_realizada and not es_clon

        if es_realizada:
            estado_display = 'realizada'
        elif es_en_proceso:
            estado_display = 'en_proceso'
        elif es_atrasada:
            estado_display = 'atrasada'
        else:
            estado_display = 'pendiente'

        # MPs necesarias (siempre se calculan para mostrar en modal, pero solo
        # se acumulan al global de faltantes si la producción aún consume stock)
        mps_nec = []
        for fi in formulas.get(producto_norm, []):
            g_total = float(fi['g_por_lote'] or 0) * lotes
            if g_total <= 0 and cant_kg_total > 0:
                g_total = (fi['porcentaje'] / 100.0) * cant_kg_total * 1000.0
            if g_total <= 0:
                continue
            mps_nec.append({
                'codigo_mp': fi['codigo_mp'],
                'nombre': fi['nombre'],
                'necesario_g': round(g_total, 2),
            })
            cod = (fi['codigo_mp'] or '').strip()
            if cod and cuenta_para_faltantes:
                consumo_mp_agregado[cod] = consumo_mp_agregado.get(cod, 0) + g_total

        # MEE necesarios · 2 caminos:
        #  A) Producto con VARIANTES (>1 presentación en producto_presentaciones):
        #     calcular envases por presentación usando ratio Shopify 90d.
        #     Cada presentación tiene su volumen_ml y su envase_codigo propios.
        #  B) Producto SIN variantes: comportamiento clásico (sku_mee_config +
        #     volumen_unitario_producto único).
        # FASE 2 · Sebastián 27-may-2026 PM · fix para LIP SERUM tipo 30ml+15ml.
        unidades_envasadas = 0
        mees_nec = []
        presentaciones_prod = presentaciones_por_producto.get(producto_norm, [])
        usar_variantes = (len(presentaciones_prod) > 1 and cant_kg_total > 0)

        if usar_variantes:
            ratios = _ratio_presentaciones(producto_norm)
            unidades_total_calc = 0
            for p in presentaciones_prod:
                ratio = ratios.get(p['codigo'], 0)
                vol_p = p['volumen_ml']
                if vol_p <= 0 or ratio <= 0:
                    continue
                # kg que va a esta presentación · luego ml · luego unidades
                kg_p = cant_kg_total * ratio
                un_p = int(round((kg_p * 1000.0) / vol_p))
                unidades_total_calc += un_p
                env_cod = p['envase_codigo']
                if env_cod and un_p > 0:
                    _desc_env = (mee_info.get(env_cod, {}) or {}).get('descripcion', env_cod) if mee_info else env_cod
                    mees_nec.append({
                        'codigo': env_cod,
                        'descripcion': _desc_env,
                        'tipo': 'envase',
                        'necesario_unidades': un_p,
                        'presentacion': p['codigo'],
                        'ratio_pct': round(ratio * 100, 1),
                    })
                    if cuenta_para_faltantes:
                        cod_norm = env_cod.strip().upper()
                        consumo_mee_agregado[cod_norm] = consumo_mee_agregado.get(cod_norm, 0) + un_p
            unidades_envasadas = unidades_total_calc
            # Además, agregar OTROS mees configurados en sku_mee_config para este
            # producto (típicamente tapas/etiquetas que comparten todas las
            # presentaciones · cant_por_unidad * unidades_envasadas total)
            for me in mee_por_producto.get(producto_norm, []):
                cant_total = unidades_envasadas * float(me['cantidad_por_unidad'] or 1)
                mees_nec.append({
                    'codigo': me['mee_codigo'],
                    'descripcion': me['descripcion'],
                    'tipo': me['tipo'],
                    'necesario_unidades': round(cant_total, 0),
                })
                cod_mee = (me['mee_codigo'] or '').strip().upper()
                if cod_mee and cuenta_para_faltantes:
                    consumo_mee_agregado[cod_mee] = consumo_mee_agregado.get(cod_mee, 0) + cant_total
        else:
            # Camino B · clásico (1 presentación o sin presentaciones definidas)
            vol_ml = volumen_por_producto.get(producto_norm, 0)
            if vol_ml > 0 and cant_kg_total > 0:
                # Aproximación: 1g ≈ 1ml para productos cosméticos (densidad ~1)
                unidades_envasadas = int(round((cant_kg_total * 1000.0) / vol_ml))
            elif cant_kg_total > 0 and cuenta_para_faltantes and mee_por_producto.get(producto_norm):
                # Producto con envases configurados pero SIN volumen_ml registrado:
                # no se puede calcular cuántos envases necesita → su faltante de
                # MEE quedaría en CERO (sub-conteo silencioso). Lo reportamos para
                # que el usuario complete el volumen.
                productos_sin_volumen.add(producto or producto_norm)

            for me in mee_por_producto.get(producto_norm, []):
                if unidades_envasadas <= 0:
                    continue
                cant_total = unidades_envasadas * float(me['cantidad_por_unidad'] or 1)
                mees_nec.append({
                    'codigo': me['mee_codigo'],
                    'descripcion': me['descripcion'],
                    'tipo': me['tipo'],
                    'necesario_unidades': round(cant_total, 0),
                })
                cod_mee = (me['mee_codigo'] or '').strip().upper()
                if cod_mee and cuenta_para_faltantes:
                    consumo_mee_agregado[cod_mee] = consumo_mee_agregado.get(cod_mee, 0) + cant_total

        producciones_out.append({
            'id': pid,
            'producto': producto,
            'fecha': fecha,
            'lotes': lotes,
            'cantidad_kg': cant_kg_total,
            'mps_necesarias': mps_nec,
            'mees_necesarios': mees_nec,
            'unidades_envasadas_estimadas': unidades_envasadas,
            # Sebastian 5-may-2026: campos para vista calendario por sala
            'area_id': area_id,
            'area_codigo': area_codigo or '',
            'area_nombre': area_nombre or '',
            'estado': estado_norm,
            'inicio_real_at': inicio_real_at or None,
            'fin_real_at': fin_real_at or None,
            # Sebastián 8-may-2026: estado_display + flags rescate
            'estado_display': estado_display,
            'realizada': es_realizada,
            'en_proceso': es_en_proceso,
            'atrasada': es_atrasada,
            'inventario_descontado_at': desc_at or None,
            'es_clon_ignorado': es_clon,
        })

    # 10. Calcular faltantes agregados
    # Lo que YA está pedido NO se vuelve a pedir: OCs activas sin recibir +
    # SOLs pendientes sin OC. Sin esto el panel muestra como "faltante" lo
    # que Catalina ya compró → sobrecompra y SOLs duplicadas en cada recarga.
    # Clave en MAYÚSCULA · MP y MEE no comparten código.
    pedido_por_codigo = {}
    try:
        for r in c.execute("""
            SELECT UPPER(TRIM(i.codigo_mp)),
                   COALESCE(SUM(i.cantidad_g - COALESCE(i.cantidad_recibida_g,0)),0)
            FROM ordenes_compra_items i
            JOIN ordenes_compra oc ON oc.numero_oc = i.numero_oc
            WHERE oc.estado IN ('Borrador','Pendiente','Revisada','Aprobada',
                                'Autorizada','Parcial','Pagada')
              AND COALESCE(TRIM(i.codigo_mp),'') != ''
            GROUP BY UPPER(TRIM(i.codigo_mp))
        """).fetchall():
            pedido_por_codigo[r[0]] = pedido_por_codigo.get(r[0], 0.0) + float(r[1] or 0)
    except sqlite3.OperationalError:
        pass
    try:
        for r in c.execute("""
            SELECT UPPER(TRIM(si.codigo_mp)), COALESCE(SUM(si.cantidad_g),0)
            FROM solicitudes_compra_items si
            JOIN solicitudes_compra s ON s.numero = si.numero
            WHERE s.estado IN ('Pendiente','En revision','Aprobada')
              AND COALESCE(s.numero_oc,'') = ''
              AND COALESCE(TRIM(si.codigo_mp),'') != ''
            GROUP BY UPPER(TRIM(si.codigo_mp))
        """).fetchall():
            pedido_por_codigo[r[0]] = pedido_por_codigo.get(r[0], 0.0) + float(r[1] or 0)
    except sqlite3.OperationalError:
        pass

    faltantes_mps = []
    for cod, g_total in consumo_mp_agregado.items():
        info = mp_info.get(cod, {'nombre': cod, 'proveedor': ''})
        # FIX 1-jun-2026 audit · lookup canónico 5-tier (antes solo id + nombre exacto
        # → MP con variante de nombre no bridgeada daba stock 0 → faltante FALSO →
        # sobre-compra · esto alimenta solicitar-faltantes-bulk que crea SOLs reales).
        s_actual = _lookup_stock_5tier(stock_mp, cod, info.get('nombre') or '')
        ya_pedido = pedido_por_codigo.get((cod or '').strip().upper(), 0.0)
        faltante = max(0.0, g_total - s_actual - ya_pedido)
        if faltante > 0:
            faltantes_mps.append({
                'codigo_mp': cod,
                'nombre': info['nombre'],
                'necesario_total_g': round(g_total, 2),
                'stock_actual_g': round(s_actual, 2),
                'ya_pedido_g': round(ya_pedido, 2),
                'faltante_g': round(faltante, 2),
                'proveedor_sugerido': info['proveedor'],
            })
    faltantes_mps.sort(key=lambda x: -x['faltante_g'])

    faltantes_mees = []
    for cod, u_total in consumo_mee_agregado.items():
        info = mee_info.get(cod, {'descripcion': cod, 'proveedor': ''})
        s_actual = float(stock_mee.get(cod, 0))
        ya_pedido = pedido_por_codigo.get((cod or '').strip().upper(), 0.0)
        faltante = max(0.0, u_total - s_actual - ya_pedido)
        if faltante > 0:
            faltantes_mees.append({
                'codigo': cod,
                'descripcion': info['descripcion'],
                'tipo': '',
                'necesario_total_u': round(u_total, 0),
                'stock_actual_u': round(s_actual, 0),
                'ya_pedido_u': round(ya_pedido, 0),
                'faltante_u': round(faltante, 0),
                'proveedor_sugerido': info['proveedor'],
            })
    faltantes_mees.sort(key=lambda x: -x['faltante_u'])

    # Proveedores únicos involucrados
    provs = {f['proveedor_sugerido'] for f in faltantes_mps if f['proveedor_sugerido']}
    provs |= {f['proveedor_sugerido'] for f in faltantes_mees if f['proveedor_sugerido']}

    # Sebastian 7-may-2026: agrupación por producto · una entry por producto
    # con TODAS sus fechas listadas, total kg, total lotes, MPs/MEEs faltantes
    # agregados. Soluciona el ruido visual de "lo mismo aparece en lunes y
    # miércoles" cuando son producciones duplicadas en `produccion_programada`.
    #
    # Detección de duplicados sospechosos: si dos producciones del MISMO
    # producto tienen exactamente mismos `lotes` y `cantidad_kg` y sus fechas
    # están dentro de 7 días, marcamos el grupo con `duplicado_sospechoso=True`
    # — Sebastián decide si limpiar via /api/programacion/limpiar-duplicados-producciones.
    grupos_por_producto = {}
    for p in producciones_out:
        key = (p.get('producto') or '').strip().upper()
        if not key:
            continue
        if key not in grupos_por_producto:
            grupos_por_producto[key] = {
                'producto': p.get('producto'),
                'fechas': [],          # lista de {fecha, pid, lotes, cantidad_kg, estado}
                'total_kg': 0.0,
                'total_lotes': 0,
                'mps_necesarias_set': {},   # codigo_mp -> {nombre, necesario_g}
                'mees_necesarios_set': {},  # codigo -> {descripcion, necesario_unidades}
                'unidades_envasadas_total': 0,
                'duplicado_sospechoso': False,
                'duplicados_detalle': [],  # entries que disparan la marca
            }
        g = grupos_por_producto[key]
        g['fechas'].append({
            'pid': p.get('id'),
            'fecha': p.get('fecha'),
            'lotes': p.get('lotes', 1),
            'cantidad_kg': p.get('cantidad_kg', 0),
            'estado': p.get('estado'),
            'area_nombre': p.get('area_nombre'),
            # Sebastián 8-may-2026: flags para que frontend distinga visualmente
            'estado_display': p.get('estado_display') or 'pendiente',
            'realizada': bool(p.get('realizada')),
            'en_proceso': bool(p.get('en_proceso')),
            'atrasada': bool(p.get('atrasada')),
            'inicio_real_at': p.get('inicio_real_at'),
            'fin_real_at': p.get('fin_real_at'),
            'inventario_descontado_at': p.get('inventario_descontado_at'),
        })
        # Solo sumar a totales lo que aún no se ha realizado · esos kg/lotes
        # representan trabajo PENDIENTE para Luis Enrique. Las realizadas se
        # ven como contexto pero no inflan la cifra "kg por hacer".
        if not p.get('realizada') and not p.get('en_proceso'):
            g['total_kg'] += float(p.get('cantidad_kg') or 0)
            g['total_lotes'] += int(p.get('lotes') or 0)
            g['unidades_envasadas_total'] += int(p.get('unidades_envasadas_estimadas') or 0)
        for m in (p.get('mps_necesarias') or []):
            cod = m.get('codigo_mp') or ''
            if not cod:
                continue
            existing = g['mps_necesarias_set'].get(cod)
            if existing:
                existing['necesario_g'] = (existing.get('necesario_g') or 0) + (m.get('necesario_g') or 0)
            else:
                g['mps_necesarias_set'][cod] = {
                    'codigo_mp': cod,
                    'nombre': m.get('nombre') or cod,
                    'necesario_g': m.get('necesario_g') or 0,
                }
        for me in (p.get('mees_necesarios') or []):
            cod = me.get('codigo') or ''
            if not cod:
                continue
            existing = g['mees_necesarios_set'].get(cod)
            if existing:
                existing['necesario_unidades'] = (existing.get('necesario_unidades') or 0) + (me.get('necesario_unidades') or 0)
            else:
                g['mees_necesarios_set'][cod] = {
                    'codigo': cod,
                    'descripcion': me.get('descripcion') or cod,
                    'tipo': me.get('tipo') or '',
                    'necesario_unidades': me.get('necesario_unidades') or 0,
                }

    # Detectar duplicados sospechosos · 2 producciones del mismo producto con
    # mismos lotes + cantidad_kg y fechas dentro de 7 días = casi seguro un
    # clon que quedó por sync mal hecho del calendar. Marcar grupo.
    from datetime import timedelta as _td, date as _date
    def _parse_iso(s):
        try:
            return _date.fromisoformat((s or '')[:10])
        except (ValueError, TypeError):
            return None
    for g in grupos_por_producto.values():
        # Sebastián 8-may-2026: solo comparar entre fechas NO-realizadas. Una
        # producción ya completada no es "clon" de la próxima programada — es
        # historia. Evita falsos positivos como "AZHC Lun 4 (realizada) clona
        # con AZHC Lun 11 (pendiente)".
        fs = [f for f in g['fechas']
              if not f.get('realizada') and not f.get('en_proceso')]
        if len(fs) <= 1:
            continue
        # Comparar pares
        for i in range(len(fs)):
            for j in range(i + 1, len(fs)):
                a, b = fs[i], fs[j]
                if (a.get('lotes') == b.get('lotes')
                        and abs(float(a.get('cantidad_kg') or 0)
                                - float(b.get('cantidad_kg') or 0)) < 0.01):
                    da, db = _parse_iso(a.get('fecha')), _parse_iso(b.get('fecha'))
                    if da and db and abs((da - db).days) <= 7:
                        g['duplicado_sospechoso'] = True
                        g['duplicados_detalle'].append({
                            'pid_a': a.get('pid'), 'fecha_a': a.get('fecha'),
                            'pid_b': b.get('pid'), 'fecha_b': b.get('fecha'),
                            'lotes': a.get('lotes'),
                            'cantidad_kg': a.get('cantidad_kg'),
                            'dias_separacion': abs((da - db).days),
                        })

    # Convertir sets a listas + ordenar
    producciones_agrupadas = []
    for g in grupos_por_producto.values():
        g['fechas'].sort(key=lambda x: x.get('fecha') or '')
        g['mps_necesarias'] = sorted(g.pop('mps_necesarias_set').values(),
                                      key=lambda m: m.get('nombre') or '')
        g['mees_necesarios'] = sorted(g.pop('mees_necesarios_set').values(),
                                       key=lambda m: m.get('descripcion') or '')
        # Faltantes agregados para este producto · join con sets globales
        mp_falt_set = {f['codigo_mp']: f for f in faltantes_mps}
        mee_falt_set = {f['codigo']: f for f in faltantes_mees}
        g['faltantes_mps_count'] = sum(1 for m in g['mps_necesarias']
                                        if m['codigo_mp'] in mp_falt_set)
        g['faltantes_mees_count'] = sum(1 for m in g['mees_necesarios']
                                         if m['codigo'] in mee_falt_set)
        producciones_agrupadas.append(g)
    # Orden: con duplicados sospechosos primero, luego por nombre
    producciones_agrupadas.sort(key=lambda x: (
        not x['duplicado_sospechoso'],
        x['producto'] or '',
    ))

    # Sebastián 8-may-2026: contadores por estado para el resumen del panel
    n_realizadas = sum(1 for p in producciones_out if p.get('realizada'))
    n_en_proceso = sum(1 for p in producciones_out if p.get('en_proceso'))
    n_atrasadas = sum(1 for p in producciones_out if p.get('atrasada'))
    n_pendientes = sum(1 for p in producciones_out
                       if not p.get('realizada') and not p.get('en_proceso')
                       and not p.get('atrasada'))

    # Sebastián 9-may-2026: contar atrasadas pendientes OCULTAS (anteriores
    # a atrasadas_cutoff) para mostrar en el toggle "Mostrar atrasadas viejas (N)".
    # Una atrasada-pendiente vieja es: estado != cancelado, sin inicio_real_at,
    # sin desc_at, fecha < atrasadas_cutoff.
    # NO se aplica past_window aquí · una basura de hace 30d sigue siendo
    # ruido oculto que el user puede querer ver/limpiar.
    n_atrasadas_ocultas = 0
    try:
        row = c.execute("""
            SELECT COUNT(*) FROM produccion_programada
            WHERE LOWER(COALESCE(estado, '')) != 'cancelado'
              AND COALESCE(inicio_real_at, '') = ''
              AND COALESCE(inventario_descontado_at, '') = ''
              AND fecha_programada < ?
        """, (atrasadas_cutoff,)).fetchone()
        n_atrasadas_ocultas = int(row[0] or 0) if row else 0
    except Exception:
        pass

    return jsonify({
        'horizonte_dias': dias,
        'pasado_dias': pasado_dias,
        'atrasadas_max_dias': atrasadas_max_dias,
        'producciones': producciones_out,
        'producciones_agrupadas': producciones_agrupadas,
        'faltantes_mps': faltantes_mps,
        'faltantes_mees': faltantes_mees,
        'productos_sin_volumen': sorted(productos_sin_volumen),
        'resumen': {
            'n_producciones': len(producciones_out),
            'n_productos_unicos': len(producciones_agrupadas),
            'n_productos_con_duplicados': sum(1 for g in producciones_agrupadas
                                                if g['duplicado_sospechoso']),
            'n_mps_faltantes': len(faltantes_mps),
            'n_mees_faltantes': len(faltantes_mees),
            'n_productos_sin_volumen': len(productos_sin_volumen),
            'n_proveedores_unicos': len(provs),
            'n_realizadas': n_realizadas,
            'n_en_proceso': n_en_proceso,
            'n_atrasadas': n_atrasadas,
            'n_pendientes': n_pendientes,
            'n_atrasadas_ocultas': n_atrasadas_ocultas,
        },
    })


@bp.route('/api/programacion/limpiar-duplicados-producciones', methods=['POST'])
def limpiar_duplicados_producciones():
    """Detecta y elimina producciones duplicadas en `produccion_programada`.

    Sebastián 7-may-2026 (Luis Enrique): el calendario mostraba la misma
    producción en lunes y miércoles porque algún sync dejó clones en la
    tabla. Este endpoint detecta pares con mismo producto + lotes +
    cantidad_kg dentro de 7 días, conserva la fecha más temprana y borra
    las posteriores.

    No toca:
      · Producciones ya descontadas (`inventario_descontado_at` IS NOT NULL)
      · Producciones canceladas/completadas
      · Producciones con `inicio_real_at` (ya empezó · es real, no clon)

    Body:
      {dry_run: bool, horizonte_dias: int}  (default dry_run=True, horizonte=30)

    Returns:
      200 {ok, dry_run, grupos_detectados, producciones_borradas,
           plan: [{producto, fechas: [{pid, fecha}]}], mensaje}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    # Solo COMPRAS+ADMIN+jefe de producción pueden limpiar
    permitidos = set(COMPRAS_USERS) | set(ADMIN_USERS)
    if user not in permitidos:
        return jsonify({'error': 'Solo Compras/Admin/Jefe Producción pueden limpiar'}), 403

    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get('dry_run', True))
    try:
        horizonte = max(7, min(int(body.get('horizonte_dias', 30)), 365))
    except (ValueError, TypeError):
        horizonte = 30

    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    fecha_cutoff = (datetime.now() + timedelta(days=horizonte)).strftime('%Y-%m-%d')

    # Cargar candidatos · producciones que NO están descontadas, ni canceladas/
    # completadas, ni ya empezadas, en el horizonte.
    # Sebastián 19-may-2026 · FIX hueco principio Fijo vs Sugerido:
    # excluir 'eos_plan' / 'eos_b2b' / 'eos_retroactivo'. Antes este endpoint
    # podía borrar Fijos (lo que Alejandro arrastró/editó o pedidos B2B)
    # si por azar coincidían producto+lotes+kg con una canónica cercana.
    rows = c.execute("""
        SELECT id, producto, fecha_programada,
               COALESCE(lotes, 1), COALESCE(cantidad_kg, 0),
               COALESCE(LOWER(estado), 'pendiente') as estado_norm,
               COALESCE(inicio_real_at, '') as inicio,
               COALESCE(inventario_descontado_at, '') as desc_at
        FROM produccion_programada
        WHERE COALESCE(inventario_descontado_at, '') = ''
          AND LOWER(COALESCE(estado, '')) NOT IN ('cancelado', 'completado')
          AND COALESCE(inicio_real_at, '') = ''
          AND COALESCE(origen, '') NOT IN ('eos_plan', 'eos_b2b', 'eos_retroactivo')
          AND fecha_programada >= ?
          AND fecha_programada <= ?
        ORDER BY producto, fecha_programada ASC
    """, (fecha_hoy, fecha_cutoff)).fetchall()

    # Agrupar por (producto_norm, lotes, cantidad_kg) y buscar pares dentro de 7d
    from datetime import date as _date
    def _iso(s):
        try:
            return _date.fromisoformat((s or '')[:10])
        except (ValueError, TypeError):
            return None

    grupos = {}
    for r in rows:
        pid, producto, fecha, lotes, cant_kg, _est, _ini, _desc = r
        key = ((producto or '').strip().upper(), int(lotes or 1),
               round(float(cant_kg or 0), 2))
        grupos.setdefault(key, []).append({'pid': pid, 'fecha': fecha,
                                            'producto': producto})

    # Para cada grupo con >1 entries, encontrar clusters de fechas <=7d apart
    a_borrar = []  # [{pid, producto, fecha}]
    plan = []      # [{producto, fechas: [{pid, fecha, accion}]}]
    for (prod_norm, lotes, cant), entries in grupos.items():
        if len(entries) < 2:
            continue
        # Ordenar por fecha
        entries.sort(key=lambda e: e.get('fecha') or '')
        # Tomar la primera como "ancla" · borrar las demás SI están dentro de
        # 7 días de la ancla
        ancla = entries[0]
        ancla_d = _iso(ancla['fecha'])
        if not ancla_d:
            continue
        clones = []
        for e in entries[1:]:
            ed = _iso(e.get('fecha'))
            if ed and abs((ed - ancla_d).days) <= 7:
                clones.append(e)
        if not clones:
            continue
        plan.append({
            'producto': ancla['producto'],
            'lotes': lotes, 'cantidad_kg': cant,
            'fechas': (
                [{'pid': ancla['pid'], 'fecha': ancla['fecha'],
                  'accion': 'KEEP (más temprana)'}]
                + [{'pid': cl['pid'], 'fecha': cl['fecha'], 'accion': 'BORRAR'}
                   for cl in clones]
            ),
        })
        for cl in clones:
            a_borrar.append({'pid': cl['pid'], 'producto': ancla['producto'],
                              'fecha': cl.get('fecha')})

    if dry_run:
        return jsonify({
            'ok': True,
            'dry_run': True,
            'grupos_detectados': len(plan),
            'producciones_a_borrar': len(a_borrar),
            'plan': plan,
            'mensaje': (f'Detectados {len(plan)} grupos con duplicados · '
                        f'{len(a_borrar)} producciones se borrarían'),
        }), 200

    if not a_borrar:
        return jsonify({
            'ok': True, 'dry_run': False,
            'grupos_detectados': 0, 'producciones_borradas': 0,
            'plan': [], 'mensaje': 'No hay duplicados que limpiar',
        }), 200

    try:
        pids = [b['pid'] for b in a_borrar]
        ph = ','.join(['?'] * len(pids))
        # Sebastián 19-may-2026 · FIX: era DELETE duro que dejaba la
        # producción sin rastro. Ahora soft-cancel con marca en observaciones
        # y mantiene el row para auditoría / recuperación. Defensa en
        # profundidad: el WHERE TAMBIÉN excluye orígenes Fijos por si el
        # SELECT cambiara en el futuro (cinturón + tirantes).
        c.execute(
            f"""UPDATE produccion_programada
                SET estado = 'cancelado',
                    observaciones = COALESCE(observaciones, '')
                      || ' · CANCELADO_LIMPIAR_DUPLICADOS_'
                      || datetime('now', '-5 hours')
                WHERE id IN ({ph})
                  AND COALESCE(origen, '') NOT IN ('eos_plan', 'eos_b2b', 'eos_retroactivo')""",
            pids,
        )
        canceladas = c.rowcount or 0
        try:
            audit_log(
                c, usuario=user,
                accion='LIMPIAR_DUPLICADOS_PRODUCCIONES',
                tabla='produccion_programada',
                registro_id='bulk',
                antes={'ids': pids},
                despues={'grupos': len(plan), 'canceladas': canceladas,
                          'horizonte_dias': horizonte},
                detalle=(f"Soft-canceló {canceladas} producciones duplicadas en "
                          f"{len(plan)} grupos · horizonte {horizonte}d · "
                          f"excluyó eos_plan/eos_b2b/eos_retroactivo"),
            )
        except Exception as e:
            log.warning('audit_log LIMPIAR_DUPLICADOS_PRODUCCIONES falló: %s', e)
        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        log.exception('limpiar_duplicados_producciones falló: %s', e)
        return jsonify({'error': str(e)}), 500

    return jsonify({
        'ok': True, 'dry_run': False,
        'grupos_detectados': len(plan),
        # mantener clave `producciones_borradas` para compat con frontend existente
        'producciones_borradas': canceladas,
        'producciones_canceladas': canceladas,
        'plan': plan,
        'mensaje': (f'✓ Canceladas {canceladas} producciones duplicadas en '
                     f'{len(plan)} grupos · horizonte {horizonte}d · '
                     f'lo Fijo NO se tocó'),
    }), 200


@bp.route('/api/abastecimiento/consumo-horizontes', methods=['GET'])
def abastecimiento_consumo_horizontes():
    """MRP por múltiples horizontes · consumo de MP/MEE según producciones Fijas
    + pedidos B2B pendientes en N horizontes (default 15/30/60/90/120/180/365 días).

    Sebastián 23-may-2026: 'abastecimiento debería ser consumo · qué se va a
    consumir según las producciones de 15, 30, 60, 90, 120, 180 y 365 días'.

    Para cada MP/MEE devuelve:
      - stock_actual + pendiente_compras (lo que ya está en cola)
      - consumo en cada horizonte (acumulativo · 15d incluido en 30d, etc.)
      - déficit por horizonte (consumo - stock - pendiente)
      - urgencia = primer horizonte donde el déficit > 0

    Reglas:
      - Solo producciones Fijas (origen IN eos_plan/eos_b2b/eos_retroactivo)
      - Excluye producciones canceladas, terminadas y ya descontadas
      - Pedidos B2B pendientes (sin producción asociada) cuentan también
      - Cubre MP (formula_items) y MEE (sku_mee_config) según `tipo` query param

    Query params:
      horizontes: CSV de int días (default '15,30,60,90,120,180,365')
      tipo: 'mp', 'mee' o 'mp,mee' (default ambos)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    # Parse horizontes
    raw_h = request.args.get('horizontes', '15,30,60,90,120,180,365')
    try:
        horizontes = sorted({
            max(1, min(int(h.strip()), 730))
            for h in raw_h.split(',') if h.strip().isdigit()
        })
    except Exception:
        horizontes = [15, 30, 60, 90, 120, 180, 365]
    if not horizontes:
        horizontes = [15, 30, 60, 90, 120, 180, 365]

    tipo = (request.args.get('tipo', 'mp,mee') or '').lower()
    incluir_mp = 'mp' in tipo
    incluir_mee = 'mee' in tipo

    # Modo de demanda · Sebastián 23-may-2026 · dual sin inflar
    # · 'comprometido' (default): TODO lo del Calendario (Fijo + Sugerida)
    # · 'run_rate': agrega proyección por velocidad de ventas Shopify
    #   (consumo proyectado = velocidad_kg_dia × h · descuenta lo ya en
    #   Calendario para evitar doble-contar)
    # FIX #3 · 23-may-2026 · "el abastecimiento debería ser tomado desde el
    # calendario donde tenemos programado todo por varios meses". Antes solo
    # consideraba origen eos_plan/b2b/retroactivo (Fijo) · ahora también
    # lee eos_canonico/auto_plan/sugerido (Sugeridas autoprogramadas).
    # Cuando el usuario fija algo Sugerido, REPROGRAMAR lo promueve a
    # eos_plan automáticamente · sin duplicar.
    # Param ?solo_fijo=1 para conservar comportamiento legacy.
    modo = (request.args.get('modo', 'comprometido') or '').lower()
    if modo not in ('comprometido', 'run_rate'):
        modo = 'comprometido'
    solo_fijo = request.args.get('solo_fijo', '').lower() in ('1','true','yes','si')

    conn = get_db()
    c = conn.cursor()
    from datetime import date, timedelta as _td
    hoy = date.today()
    cutoff_max = (hoy + _td(days=max(horizontes))).isoformat()
    hoy_iso = hoy.isoformat()

    # 1. Producciones del Calendario · Fijo + Sugerida (a menos que solo_fijo)
    # exclude: cancelado, completado y ya descontadas (su MP ya bajó del stock)
    if solo_fijo:
        origenes_in = ('eos_plan', 'eos_b2b', 'eos_retroactivo')
    else:
        origenes_in = ('eos_plan', 'eos_b2b', 'eos_retroactivo',
                        'eos_canonico', 'auto_plan', 'sugerido')
    placeholders = ','.join(['?'] * len(origenes_in))
    # Sebastián 25-may-2026 PM · incluir envase_codigo_override (mig 184).
    # Si el lote tiene override no vacío, el cálculo MEE prioriza ese envase
    # sobre el default del sku_mee_config. Fallback si mig 184 no aplicada.
    try:
        prod_rows = c.execute(f"""
            SELECT pp.id, pp.producto, pp.fecha_programada,
                   COALESCE(pp.lotes, 1), COALESCE(pp.cantidad_kg, 0),
                   COALESCE(fh.lote_size_kg, 0),
                   COALESCE(pp.origen, ''),
                   COALESCE(pp.envase_codigo_override, '')
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh
                   ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE COALESCE(pp.origen,'') IN ({placeholders})
              AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado','esperando_recurso')
              AND COALESCE(pp.inventario_descontado_at,'') = ''
              AND pp.fecha_programada >= ?
              AND pp.fecha_programada <= ?
            ORDER BY pp.fecha_programada ASC
        """, origenes_in + (hoy_iso, cutoff_max)).fetchall()
        _has_env_ovr_calc = True
    except Exception:
        prod_rows = c.execute(f"""
            SELECT pp.id, pp.producto, pp.fecha_programada,
                   COALESCE(pp.lotes, 1), COALESCE(pp.cantidad_kg, 0),
                   COALESCE(fh.lote_size_kg, 0),
                   COALESCE(pp.origen, '')
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh
                   ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE COALESCE(pp.origen,'') IN ({placeholders})
              AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado','esperando_recurso')
              AND COALESCE(pp.inventario_descontado_at,'') = ''
              AND pp.fecha_programada >= ?
              AND pp.fecha_programada <= ?
            ORDER BY pp.fecha_programada ASC
        """, origenes_in + (hoy_iso, cutoff_max)).fetchall()
        _has_env_ovr_calc = False
    # FIX AUDIT 24-may-2026 noche · agente cazó: 'esperando_recurso' es
    # un lote pausado por falta de MP. Si lo cuento como consumo, el
    # déficit que reporta Abastecimiento es CIRCULAR: el lote consume
    # MP que no tengo → el lote queda pausado → pero sigue contando.
    # Ahora excluido junto con cancelado/completado.

    # 2. Pedidos B2B PENDIENTES (sin producción Fija eos_b2b aún)
    # AUDITORÍA-FIX 23-may-2026 · agente cazó 3 bugs:
    #   · 'programado' no existe en CHECK · era 'pendiente','confirmado','en_produccion'
    #   · 'confirmado'/'en_produccion' suelen ya tener Fija eos_b2b · doble-conteo
    #   · solución conservadora: solo 'pendiente' (lo no convertido a Fija aún)
    # FIX AUDIT 24-may-2026 noche · LEFT JOIN para excluir pedidos que YA
    # fueron integrados a un lote en `pedidos_b2b_lote` (mig 171). Antes
    # si la integración pasó pero el estado_pedido seguía 'pendiente' (race
    # condition o fallo parcial), el pedido se contaba en (8a) vía
    # produccion_programada Y en (8b) vía pedidos_b2b → doble cuenta.
    b2b_rows = []
    try:
        b2b_rows = c.execute("""
            SELECT pb.id, pb.cliente_id, pb.cliente_nombre, pb.producto_nombre,
                   COALESCE(pb.cantidad_uds,0), COALESCE(pb.ml_unidad,30),
                   COALESCE(pb.fecha_estimada,'')
            FROM pedidos_b2b pb
            LEFT JOIN pedidos_b2b_lote pbl ON pbl.pedido_b2b_id = pb.id
            WHERE LOWER(COALESCE(pb.estado,'pendiente')) = 'pendiente'
              AND COALESCE(pb.fecha_estimada,'') >= ?
              AND COALESCE(pb.fecha_estimada,'') <= ?
              AND pbl.id IS NULL
        """, (hoy_iso, cutoff_max)).fetchall()
    except sqlite3.OperationalError:
        b2b_rows = []

    # 3. Fórmulas por producto (g por lote_kg)
    # AUDITORÍA-FIX 23-may-2026 · agente cazó 2 bugs:
    #   · No filtraba formula_headers.activo=1 · fórmulas archivadas inflaban
    #   · codigo_mp se guardaba raw (sin UPPER+TRIM) · luego mismatch con
    #     pendientes_mp (UPPER+TRIM) y stock_mp (en _get_mp_stock guarda raw
    #     pero hay name-variant lookup · mejor normalizar consistente)
    # FIX 24-may-2026 noche · Sebastián: 'sigue mal · no me cuadra'.
    # SQL UPPER(TRIM(...)) NO colapsa espacios DOBLES interiores. Los seeds
    # de mig_127 tienen productos con dobles espacios ('EMULSION HIDRATANTE
    # _ B3+BHA') que no matchean con la producción ('EMULSION HIDRATANTE
    # B3+BHA' un solo espacio). Resultado: lote_size=0 → producción saltada
    # del cálculo. Ahora normalizamos en Python · colapsa espacios múltiples.
    def _norm_prod(s):
        return ' '.join((s or '').strip().upper().split())
    formulas = {}
    # Pre-cargar lote_size_kg por producto (también para B2B · evita N+1)
    lote_size_por_prod = {}
    # FIX 1-jun-2026 audit Abastecimiento (P2) · (a) JOIN normalizado UPPER(TRIM)
    # → antes la igualdad exacta perdía TODOS los items de un producto si el
    # nombre difería en case/espacios entre formula_items y formula_headers
    # (sub-conteo de demanda). (b) dedup por (producto, codigo_mp) → formula_items
    # no tiene UNIQUE, los duplicados sumaban el MP dos veces (sobre-conteo).
    _seen_fi = set()
    for r in c.execute("""
        SELECT fi.producto_nombre, fi.material_id,
               COALESCE(fi.material_nombre,''),
               COALESCE(fi.porcentaje,0), COALESCE(fi.cantidad_g_por_lote,0),
               COALESCE(fh.lote_size_kg, 0)
        FROM formula_items fi
        JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(fi.producto_nombre))
        WHERE fi.material_id IS NOT NULL AND TRIM(fi.material_id) != ''
          AND COALESCE(fh.activo,1) = 1
    """).fetchall():
        cm_norm = str(r[1] or '').strip().upper()
        if not cm_norm:
            continue
        prod_key = _norm_prod(r[0])
        _dk = (prod_key, cm_norm)
        if _dk in _seen_fi:
            continue  # duplicado de formula_items · no doble-contar
        _seen_fi.add(_dk)
        formulas.setdefault(prod_key, []).append({
            'codigo_mp': cm_norm,
            'nombre': r[2],
            'pct': float(r[3] or 0),
            'g_por_lote': float(r[4] or 0),
        })
        if prod_key not in lote_size_por_prod:
            lote_size_por_prod[prod_key] = float(r[5] or 0)

    # 4. MEE por producto (volumen-aware)
    # AUDITORÍA-FIX 23-may-2026 · agente P0-2 · sku_mee_config.sku_codigo es
    # un SKU comercial (ej 'SAH-30') pero el endpoint indexaba por
    # producto_norm (ej 'SUERO HIDRATANTE AH 1.5%') → match imposible · TODO
    # MEE devolvía 0 silente · ahora resuelve producto → SKUs (via
    # sku_producto_map) → MEE configs (via sku_mee_config)
    mee_por_producto = {}    # producto_nombre_upper → [{codigo, cant_x_uds, tipo}]
    volumen_por_producto = {}
    if incluir_mee:
        # Primero: SKU → MEE configs
        sku_to_mee = {}
        try:
            for r in c.execute("""
                SELECT UPPER(TRIM(s.sku_codigo)), s.mee_codigo,
                       COALESCE(s.cantidad_por_unidad,1),
                       COALESCE(s.componente_tipo,'envase')
                FROM sku_mee_config s
                WHERE COALESCE(s.aplica,1)=1
            """).fetchall():
                sku_to_mee.setdefault(r[0], []).append({
                    'codigo': r[1],
                    'cant_x_uds': float(r[2] or 1),
                    'tipo': r[3] or 'envase',
                })
        except sqlite3.OperationalError:
            pass
        # Segundo: SKU → producto vía sku_producto_map (sí existe)
        try:
            for r in c.execute("""
                SELECT UPPER(TRIM(sku)), UPPER(TRIM(producto_nombre))
                FROM sku_producto_map
                WHERE COALESCE(activo,1)=1
                  AND producto_nombre IS NOT NULL AND TRIM(producto_nombre) != ''
            """).fetchall():
                sku_u, prod_u = r[0], r[1]
                if sku_u in sku_to_mee:
                    for me in sku_to_mee[sku_u]:
                        # Cada producto agrega TODOS los MEEs de TODOS sus SKUs
                        mee_por_producto.setdefault(prod_u, []).append(me)
        except sqlite3.OperationalError:
            pass
        # Fallback: si sku_codigo coincide directo con producto_nombre (caso edge)
        for sku_u, mees in sku_to_mee.items():
            if sku_u not in mee_por_producto:
                mee_por_producto[sku_u] = mees
        try:
            for r in c.execute("""
                SELECT UPPER(TRIM(producto_nombre)), COALESCE(volumen_ml,0)
                FROM volumen_unitario_producto WHERE COALESCE(activo,1)=1
            """).fetchall():
                volumen_por_producto[r[0]] = float(r[1] or 0)
        except sqlite3.OperationalError:
            pass

    # 5. Stock canonical + info MP (con lead time)
    # AUDITORÍA-FIX 23-may-2026 · Sebastián · "centro de solicitudes" pidió
    # mostrar lead time para que se vea cuándo pedir antes de que se acabe
    stock_mp = _get_mp_stock(conn)
    mp_info = {}
    try:
        for r in c.execute("""
            SELECT mm.codigo_mp,
                   COALESCE(mm.nombre_comercial, mm.nombre_inci, mm.codigo_mp),
                   COALESCE(NULLIF(TRIM(mlt.proveedor_principal),''),
                            mm.proveedor, ''),
                   COALESCE(mlt.lead_time_dias, 14),
                   COALESCE(mlt.buffer_dias, 30),
                   COALESCE(mm.nombre_inci,'')
            FROM maestro_mps mm
            LEFT JOIN mp_lead_time_config mlt ON mlt.material_id = mm.codigo_mp
            WHERE COALESCE(mm.activo,1)=1
        """).fetchall():
            mp_info[r[0]] = {
                'nombre': r[1] or r[0],
                'proveedor': (r[2] or '').strip(),
                'lead_time_dias': int(r[3] or 14),
                'buffer_dias': int(r[4] or 30),
                'nombre_inci': (r[5] or '').strip(),
            }
    except sqlite3.OperationalError:
        pass

    # 6. Stock + info MEE
    # P0-5 23-may-PM · auditoría · antes leía stock_actual directo
    # (cache puede drift) · ahora usa _get_mee_stock canónico.
    # Info adicional (descripcion, proveedor) sigue de maestro_mee.
    stock_mee = {}
    mee_info = {}
    if incluir_mee:
        try:
            stock_mee = _get_mee_stock(conn)
            for r in c.execute("""
                SELECT codigo, COALESCE(descripcion,''),
                       COALESCE(proveedor,'')
                FROM maestro_mee
            """).fetchall():
                k = str(r[0]).strip().upper()
                mee_info[k] = {'descripcion': r[1] or r[0], 'proveedor': r[2] or ''}
        except sqlite3.OperationalError:
            pass

    # 7. Pendientes en compras (bulk · dict)
    # AUDITORÍA-FIX 23-may-2026 · P1-3 · antes pendientes_mee hardcoded en 0
    # · ahora consulta separada filtrando por sc.categoria IN ('Empaque',
    # 'Material de Empaque') para separar MP de MEE · evita sobre-pedir
    pendientes_mp = {}
    pendientes_mee = {}
    try:
        for cm, gp in c.execute("""
            SELECT UPPER(TRIM(sci.codigo_mp)),
                   COALESCE(SUM(sci.cantidad_g),0)
            FROM solicitudes_compra_items sci
            JOIN solicitudes_compra sc ON sc.numero = sci.numero
            WHERE sc.estado IN ('Pendiente','En revision','Aprobada')
              AND COALESCE(sc.numero_oc,'')=''
              AND sci.codigo_mp IS NOT NULL AND TRIM(sci.codigo_mp) != ''
              AND COALESCE(sc.categoria,'') NOT IN ('Empaque','Material de Empaque')
            GROUP BY UPPER(TRIM(sci.codigo_mp))
        """).fetchall():
            pendientes_mp[cm] = float(gp or 0)
    except Exception:
        pass
    try:
        for cm, gp in c.execute("""
            SELECT UPPER(TRIM(oci.codigo_mp)),
                   COALESCE(SUM(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g,0)),0)
            FROM ordenes_compra_items oci
            JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
            LEFT JOIN solicitudes_compra sc ON sc.numero_oc = oc.numero_oc
            WHERE oc.estado IN ('Borrador','Revisada','Autorizada','Parcial')
              AND oci.codigo_mp IS NOT NULL AND TRIM(oci.codigo_mp) != ''
              AND COALESCE(sc.categoria,'') NOT IN ('Empaque','Material de Empaque')
            GROUP BY UPPER(TRIM(oci.codigo_mp))
        """).fetchall():
            k = cm
            pendientes_mp[k] = pendientes_mp.get(k, 0.0) + float(gp or 0)
    except Exception:
        pass
    # AUDITORÍA-FIX 23-may-2026 · Sebastián · detalle SOL/OC en curso por
    # código para mostrar badge "Pendiente SOL-2026-0042" en cada fila
    solicitudes_en_curso = {}  # codigo_upper → [{tipo, numero, estado, cantidad}]
    try:
        for r in c.execute("""
            SELECT UPPER(TRIM(sci.codigo_mp)), sc.numero, sc.estado,
                   COALESCE(sci.cantidad_g, 0), COALESCE(sc.categoria,'')
            FROM solicitudes_compra_items sci
            JOIN solicitudes_compra sc ON sc.numero = sci.numero
            WHERE sc.estado IN ('Pendiente','En revision','Aprobada')
              AND COALESCE(sc.numero_oc,'')=''
              AND sci.codigo_mp IS NOT NULL AND TRIM(sci.codigo_mp) != ''
        """).fetchall():
            solicitudes_en_curso.setdefault(r[0], []).append({
                'tipo': 'SOL',
                'numero': r[1],
                'estado': r[2],
                'cantidad': float(r[3] or 0),
                'categoria': r[4],
            })
    except Exception:
        pass
    try:
        for r in c.execute("""
            SELECT UPPER(TRIM(oci.codigo_mp)), oc.numero_oc, oc.estado,
                   COALESCE(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g,0), 0),
                   COALESCE(sc.categoria,'')
            FROM ordenes_compra_items oci
            JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
            LEFT JOIN solicitudes_compra sc ON sc.numero_oc = oc.numero_oc
            WHERE oc.estado IN ('Borrador','Revisada','Autorizada','Parcial')
              AND oci.codigo_mp IS NOT NULL AND TRIM(oci.codigo_mp) != ''
        """).fetchall():
            solicitudes_en_curso.setdefault(r[0], []).append({
                'tipo': 'OC',
                'numero': r[1],
                'estado': r[2],
                'cantidad': float(r[3] or 0),
                'categoria': r[4],
            })
    except Exception:
        pass

    # MEE pendientes · misma query filtrando categoria de Empaque
    if incluir_mee:
        try:
            for cm, up in c.execute("""
                SELECT UPPER(TRIM(sci.codigo_mp)),
                       COALESCE(SUM(sci.cantidad_g),0)
                FROM solicitudes_compra_items sci
                JOIN solicitudes_compra sc ON sc.numero = sci.numero
                WHERE sc.estado IN ('Pendiente','En revision','Aprobada')
                  AND COALESCE(sc.numero_oc,'')=''
                  AND sci.codigo_mp IS NOT NULL AND TRIM(sci.codigo_mp) != ''
                  AND COALESCE(sc.categoria,'') IN ('Empaque','Material de Empaque')
                GROUP BY UPPER(TRIM(sci.codigo_mp))
            """).fetchall():
                pendientes_mee[cm] = float(up or 0)
        except Exception:
            pass
        try:
            for cm, up in c.execute("""
                SELECT UPPER(TRIM(oci.codigo_mp)),
                       COALESCE(SUM(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g,0)),0)
                FROM ordenes_compra_items oci
                JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
                LEFT JOIN solicitudes_compra sc ON sc.numero_oc = oc.numero_oc
                WHERE oc.estado IN ('Borrador','Revisada','Autorizada','Parcial')
                  AND oci.codigo_mp IS NOT NULL AND TRIM(oci.codigo_mp) != ''
                  AND COALESCE(sc.categoria,'') IN ('Empaque','Material de Empaque')
                GROUP BY UPPER(TRIM(oci.codigo_mp))
            """).fetchall():
                k = cm
                pendientes_mee[k] = pendientes_mee.get(k, 0.0) + float(up or 0)
        except Exception:
            pass

    # 8. Acumular consumo por (codigo, horizonte)
    # consumo[codigo_mp][h] = gramos · acumulativo (15d ⊂ 30d ⊂ 60d ...)
    consumo_mp = {}    # codigo_mp -> {h: g}
    consumo_mee = {}   # codigo_mee -> {h: uds}
    _zero_mp = {h: 0.0 for h in horizontes}

    def _agregar_consumo_mp(codigo, gramos, dias_hasta):
        if not codigo or gramos <= 0:
            return
        d = consumo_mp.setdefault(codigo, dict(_zero_mp))
        for h in horizontes:
            if dias_hasta <= h:
                d[h] = d.get(h, 0.0) + gramos

    def _agregar_consumo_mee(codigo, uds, dias_hasta):
        if not codigo or uds <= 0:
            return
        d = consumo_mee.setdefault(codigo, dict(_zero_mp))
        for h in horizontes:
            if dias_hasta <= h:
                d[h] = d.get(h, 0.0) + uds

    productos_sin_lote_size = set()  # AUDITORÍA P1-5 · reporte en respuesta
    productos_multi_volumen = set()  # AUDIT 24-may noche · MEE warning

    # FIX AUDIT 24-may-2026 noche · B2B envase custom y multi-volumen.
    # Bulk-cargar aportes B2B con envase específico para que el descuento
    # MEE respete el envase del cliente en lugar del default del producto.
    b2b_envases_por_lote = {}  # lote_id -> [{envase_codigo, uds, cliente}]
    if incluir_mee:
        prod_ids = [r[0] for r in prod_rows]
        if prod_ids:
            try:
                ph_pids = ','.join(['?'] * len(prod_ids))
                for ar in c.execute(
                    f"""SELECT lote_produccion_id, envase_codigo,
                              COALESCE(unidades_aporte, 0),
                              COALESCE(cliente_nombre, '')
                       FROM pedidos_b2b_lote
                       WHERE lote_produccion_id IN ({ph_pids})
                         AND COALESCE(envase_codigo, '') != ''""",
                    prod_ids,
                ).fetchall():
                    b2b_envases_por_lote.setdefault(ar[0], []).append({
                        'envase': (ar[1] or '').strip().upper(),
                        'uds': int(ar[2] or 0),
                        'cliente': ar[3] or '',
                    })
            except Exception:
                pass  # mig 171/172 no aplicada

    # FIX AUDIT 24-may-2026 noche · multi-volumen warning. Detectar
    # productos con >1 SKU con volúmenes distintos en producto_presentaciones
    # (e.g. BLUSH BALM tonos 6g + BBM MINI 15ml regalo). El cálculo actual
    # usa UN volumen ponderado · si los volúmenes son muy distintos el
    # cálculo subestima.
    multi_vol_check = {}  # producto_norm -> set(volúmenes distintos)
    if incluir_mee:
        try:
            for r in c.execute("""
                SELECT UPPER(TRIM(producto_nombre)), COALESCE(factor_g_por_unidad,0)
                FROM producto_presentaciones
                WHERE COALESCE(activo,1) = 1
                  AND COALESCE(factor_g_por_unidad,0) > 0
            """).fetchall():
                multi_vol_check.setdefault(r[0], set()).add(round(float(r[1]), 1))
        except sqlite3.OperationalError:
            pass

    # 8a. Producciones del Calendario (Fijo + Sugerida desde FIX #3 23-may)
    # FIX 24-may-2026 noche · normalizar espacios para que matchee con
    # _norm_prod usado en el diccionario formulas (colapsa espacios dobles).
    matched_lotes = 0
    sin_formula_lotes = []
    for _pr in prod_rows:
        # Unpacking flexible · si trae override (mig 184 OK) son 8 cols,
        # si no (fallback) son 7.
        if len(_pr) >= 8:
            pid, producto, fecha, lotes, cant_kg_expl, lote_size, _origen, _env_ovr = _pr
        else:
            pid, producto, fecha, lotes, cant_kg_expl, lote_size, _origen = _pr
            _env_ovr = ''
        _env_ovr = (_env_ovr or '').strip().upper()
        producto_norm = _norm_prod(producto)
        # Si el LEFT JOIN SQL devolvió lote_size=0 (por mismatch de espacios),
        # tomar el lote_size del diccionario formulas (que sí está normalizado).
        if not lote_size or float(lote_size) <= 0:
            lote_size = lote_size_por_prod.get(producto_norm, 0)
        try:
            dias_hasta = (date.fromisoformat(str(fecha)[:10]) - hoy).days
        except Exception:
            dias_hasta = 0
        if dias_hasta < 0:
            dias_hasta = 0
        cant_kg = float(cant_kg_expl or 0) or (int(lotes or 1) * float(lote_size or 0))
        if cant_kg <= 0:
            continue
        items = formulas.get(producto_norm) or []
        if items:
            matched_lotes += 1
        else:
            sin_formula_lotes.append(producto_norm)
        # FIX 24-may-2026 noche · Sebastián cazó bug raíz: en los seeds,
        # cantidad_g_por_lote se cargó IGUAL al porcentaje (typo masivo, 674
        # items). Resultado: el cálculo g_por_lote × cant_kg/lote_size daba
        # 2000× subestimado. Solución: priorizar porcentaje cuando exista
        # (más confiable matemáticamente · % × cant_kg × 1000). cantidad_g_
        # por_lote queda como FALLBACK si el % está vacío.
        # MP consumption: necesario_g por kg producido.
        # PRIORIZAR PORCENTAJE (más confiable que el seed roto de g_por_lote).
        for it in items:
            if it['pct'] > 0:
                gramos = (it['pct'] / 100.0) * cant_kg * 1000.0
            elif it['g_por_lote'] > 0 and lote_size and float(lote_size) > 0:
                gramos = it['g_por_lote'] * (cant_kg / float(lote_size))
            elif it['g_por_lote'] > 0:
                productos_sin_lote_size.add(producto_norm)
                continue
            else:
                continue
            if incluir_mp:
                _agregar_consumo_mp(it['codigo_mp'], gramos, dias_hasta)
        # MEE: requiere volumen por unidad
        if incluir_mee:
            vol = volumen_por_producto.get(producto_norm, 0.0)
            # FIX AUDIT 24-may-2026 noche · warning multi-volumen.
            vols_set = multi_vol_check.get(producto_norm, set())
            if len(vols_set) > 1:
                productos_multi_volumen.add(producto_norm)
            if vol > 0:
                uds_envasadas_total = (cant_kg * 1000.0) / vol
                # FIX AUDIT 24-may-2026 noche · B2B envase custom split.
                # Si este lote tiene aportes B2B con envase_codigo específico,
                # restar esas uds del descuento default y sumarlas al envase
                # B2B en pedidos_b2b_lote (e.g. ENV-500-FB para Fernando).
                aportes_custom = b2b_envases_por_lote.get(pid, [])
                uds_b2b_custom_total = sum(a['uds'] for a in aportes_custom)
                for a in aportes_custom:
                    _agregar_consumo_mee(a['envase'], a['uds'], dias_hasta)
                # Buscar MEE registrados por producto/sku · usar producto_norm
                # como key (sku_mee_config.sku_codigo) o fallback al codigo_pt
                mee_items = mee_por_producto.get(producto_norm, [])
                # Sebastián 25-may-2026 PM · si el lote tiene envase_codigo_override,
                # ese envase REEMPLAZA al envase default del producto (sku_mee_config).
                # Las tapas/etiquetas/sub-componentes siguen iguales · solo cambia
                # el item de tipo 'envase/frasco/recipiente'.
                for me in mee_items:
                    es_envase = (me.get('tipo') or 'envase').lower() in ('envase', 'frasco', 'recipiente')
                    cod_mee_efectivo = (me['codigo'] or '').strip().upper()
                    if es_envase and _env_ovr:
                        # Override · usa el envase elegido manualmente en lugar del default
                        cod_mee_efectivo = _env_ovr
                    if es_envase and uds_b2b_custom_total > 0:
                        uds_efectivas = max(uds_envasadas_total - uds_b2b_custom_total, 0)
                    else:
                        uds_efectivas = uds_envasadas_total
                    _agregar_consumo_mee(
                        cod_mee_efectivo,
                        uds_efectivas * me['cant_x_uds'],
                        dias_hasta,
                    )
                # Caso edge: lote tiene override pero NO hay sku_mee_config
                # (producto sin envase default) · agregar el override directo
                # con uds_envasadas_total (descontando B2B custom si lo hay)
                if _env_ovr and not any(
                    (m.get('tipo') or 'envase').lower() in ('envase','frasco','recipiente')
                    for m in mee_items):
                    uds_eff = max(uds_envasadas_total - uds_b2b_custom_total, 0)
                    _agregar_consumo_mee(_env_ovr, uds_eff, dias_hasta)

    # 8b. Pedidos B2B pendientes (sin producción Fija aún)
    for bid, cli_id, cli_nom, prod_nom, cant_uds, ml_u, f_est in b2b_rows:
        producto_norm = _norm_prod(prod_nom)
        # AUDITORÍA-FIX 23-may-2026 · P1-4 · f_est vacío caía a dias_hasta=0
        # inflaba horizonte 15d con pedidos sin fecha · ahora skip si no hay
        if not f_est:
            continue
        try:
            dias_hasta = (date.fromisoformat(str(f_est)[:10]) - hoy).days
        except Exception:
            continue
        if dias_hasta < 0:
            dias_hasta = 0
        # Convertir uds × ml a kg
        cant_kg = (float(cant_uds or 0) * float(ml_u or 30)) / 1000.0
        if cant_kg <= 0:
            continue
        items = formulas.get(producto_norm) or []
        # AUDITORÍA-FIX P1-2 · usar lote_size pre-cargado (no N+1 query)
        lote_size_b2b = lote_size_por_prod.get(producto_norm, 0.0)
        for it in items:
            # FIX 24-may noche · priorizar porcentaje (seed roto en g_por_lote)
            if it['pct'] > 0:
                gramos = (it['pct'] / 100.0) * cant_kg * 1000.0
            elif it['g_por_lote'] > 0 and lote_size_b2b > 0:
                gramos = it['g_por_lote'] * (cant_kg / lote_size_b2b)
            elif it['g_por_lote'] > 0:
                productos_sin_lote_size.add(producto_norm)
                continue
            else:
                continue
            if incluir_mp:
                _agregar_consumo_mp(it['codigo_mp'], gramos, dias_hasta)
        if incluir_mee:
            uds_envasadas = float(cant_uds or 0)
            for me in mee_por_producto.get(producto_norm, []):
                _agregar_consumo_mee(
                    (me['codigo'] or '').strip().upper(),
                    uds_envasadas * me['cant_x_uds'],
                    dias_hasta,
                )

    # 8c. RUN-RATE · proyección por velocidad de ventas Shopify
    # Sebastián 23-may-2026 · modo dual sin inflar · cuando hay pocas Fijas
    # (típicamente 2 semanas), completar con la demanda proyectada por las
    # ventas reales · NO usa sugerencias IA · solo datos históricos de ventas
    if modo == 'run_rate':
        from datetime import datetime as _dt_rr
        # Cargar velocidad por SKU últimos 60d desde animus_shopify_orders
        # (reutiliza misma fuente que /api/plan/necesidades)
        vel_60d_iso = (_dt_rr.utcnow() - _td(hours=5) - _td(days=60)).strftime('%Y-%m-%d')
        ventas_60d_sku = {}
        try:
            for r in c.execute("""
                SELECT sku_items FROM animus_shopify_orders
                WHERE creado_en >= ?
                  AND sku_items IS NOT NULL AND sku_items != ''
            """, (vel_60d_iso + 'T00:00:00',)).fetchall():
                try:
                    items = json.loads(r[0]) if r[0] else []
                except Exception:
                    continue
                for it in items:
                    sku = (it.get('sku') or '').strip().upper()
                    if not sku:
                        continue
                    qty = int(it.get('qty') or it.get('cantidad') or it.get('quantity') or 0)
                    if qty > 0:
                        ventas_60d_sku[sku] = ventas_60d_sku.get(sku, 0) + qty
        except Exception:
            pass
        # Mapeo SKU → producto + ml_unidad
        sku_to_prod_rr = {}
        try:
            for r in c.execute("""
                SELECT sku, producto_nombre FROM sku_producto_map
                WHERE COALESCE(activo,1)=1
                  AND producto_nombre IS NOT NULL AND TRIM(producto_nombre)!=''
            """).fetchall():
                sku_to_prod_rr[(r[0] or '').strip().upper()] = (r[1] or '').strip()
        except Exception:
            pass
        # Inferir ml_unidad por producto · reutiliza helper de plan.py
        try:
            from blueprints.plan import _inferir_ml_presentacion as _inf_ml
        except Exception:
            _inf_ml = lambda nombre: 30.0
        # Agregar velocidad por producto
        vel_kg_dia_por_prod = {}
        for sku, qty_60d in ventas_60d_sku.items():
            prod = sku_to_prod_rr.get(sku)
            if not prod:
                continue
            vel_uds_dia = qty_60d / 60.0
            ml = _inf_ml(prod)
            kg_dia = (vel_uds_dia * ml) / 1000.0
            # FIX 1-jun-2026 audit (P0) · usar _norm_prod (colapsa dobles espacios)
            # igual que la clave de `formulas` · antes .upper().strip() no matcheaba
            # y la proyección run-rate de esos productos caía a 0 (sub-conteo).
            prod_up = _norm_prod(prod)
            vel_kg_dia_por_prod[prod_up] = vel_kg_dia_por_prod.get(prod_up, 0) + kg_dia
        # Pre-computar kg_fijo_por_prod_por_horizonte (lo ya programado Fijo)
        # + kg_b2b_por_prod_por_horizonte para descontar del run-rate y
        # evitar doble-conteo (AUDITORÍA P2 · agente cazó este caso)
        kg_fijo_acum_por_prod = {}  # prod_up → {h: kg_acumulado}
        for _pr2 in prod_rows:
            # Unpacking flexible · soporta filas con o sin envase_codigo_override
            if len(_pr2) >= 8:
                pid, producto, fecha, lotes, cant_kg_expl, lote_size, _origen2, _ = _pr2
            else:
                pid, producto, fecha, lotes, cant_kg_expl, lote_size, _origen2 = _pr2
            producto_norm = _norm_prod(producto)  # match con clave de formulas/run-rate
            try:
                dias_hasta_f = (date.fromisoformat(str(fecha)[:10]) - hoy).days
            except Exception:
                dias_hasta_f = 0
            if dias_hasta_f < 0:
                dias_hasta_f = 0
            cant_kg_f = float(cant_kg_expl or 0) or (int(lotes or 1) * float(lote_size or 0))
            if cant_kg_f <= 0:
                continue
            d_fijo = kg_fijo_acum_por_prod.setdefault(producto_norm, dict(_zero_mp))
            for h in horizontes:
                if dias_hasta_f <= h:
                    d_fijo[h] = d_fijo.get(h, 0.0) + cant_kg_f
        # B2B también debe descontar del run-rate (los pedidos B2B ya
        # representan demanda que el run-rate proyectaría · doble-cuento si no)
        for bid, cli_id, cli_nom, prod_nom, cant_uds, ml_u, f_est in b2b_rows:
            if not f_est:
                continue
            producto_norm_b = _norm_prod(prod_nom)  # match con clave de formulas/run-rate
            try:
                dias_hasta_b = (date.fromisoformat(str(f_est)[:10]) - hoy).days
            except Exception:
                continue
            if dias_hasta_b < 0:
                dias_hasta_b = 0
            cant_kg_b = (float(cant_uds or 0) * float(ml_u or 30)) / 1000.0
            if cant_kg_b <= 0:
                continue
            d_b2b = kg_fijo_acum_por_prod.setdefault(producto_norm_b, dict(_zero_mp))
            for h in horizontes:
                if dias_hasta_b <= h:
                    d_b2b[h] = d_b2b.get(h, 0.0) + cant_kg_b
        # Agregar run-rate al consumo (solo el delta sobre Fijas)
        for prod_up, kg_dia in vel_kg_dia_por_prod.items():
            if kg_dia <= 0:
                continue
            items = formulas.get(prod_up) or []
            lote_size_rr = 0.0
            try:
                r_ls = c.execute(
                    "SELECT COALESCE(lote_size_kg,0) FROM formula_headers "
                    "WHERE UPPER(TRIM(producto_nombre))=?",
                    (prod_up,)
                ).fetchone()
                if r_ls:
                    lote_size_rr = float(r_ls[0] or 0)
            except Exception:
                pass
            fijo_acum = kg_fijo_acum_por_prod.get(prod_up, _zero_mp)
            for h in horizontes:
                kg_proyectado = kg_dia * h
                kg_ya_fijo = fijo_acum.get(h, 0.0)
                kg_delta = max(kg_proyectado - kg_ya_fijo, 0.0)
                if kg_delta <= 0.01:
                    continue
                # Explotar fórmula sobre el delta · FIX 1-jun-2026 (P1) · priorizar
                # pct ANTES que g_por_lote (igual que 8a/8b) · el seed de
                # cantidad_g_por_lote está roto (cargado = al porcentaje) → usarlo
                # primero daba escala equivocada (~hasta 2000×).
                for it in items:
                    if it['pct'] > 0:
                        gramos = (it['pct'] / 100.0) * kg_delta * 1000.0
                    elif it['g_por_lote'] > 0 and lote_size_rr > 0:
                        gramos = it['g_por_lote'] * (kg_delta / lote_size_rr)
                    else:
                        continue
                    if incluir_mp and gramos > 0:
                        # Agregar directamente al horizonte (no acumulativo
                        # porque kg_proyectado ya es acumulado hasta h)
                        d = consumo_mp.setdefault(it['codigo_mp'], dict(_zero_mp))
                        d[h] = d.get(h, 0.0) + gramos
                if incluir_mee:
                    mee_items = mee_por_producto.get(prod_up, [])
                    if mee_items:
                        # AUDITORÍA-FIX 23-may-2026 · P2 · antes heurística
                        # convoluta con .title() · ahora usa volumen_por_producto
                        # canonical primero · fallback _inf_ml
                        ml = volumen_por_producto.get(prod_up, 0.0) or _inf_ml(prod_up)
                        if ml > 0:
                            uds_envasadas = (kg_delta * 1000.0) / ml
                            for me in mee_items:
                                cod_mee = (me['codigo'] or '').strip().upper()
                                if not cod_mee:
                                    continue
                                uds = uds_envasadas * me['cant_x_uds']
                                d = consumo_mee.setdefault(cod_mee, dict(_zero_mp))
                                d[h] = d.get(h, 0.0) + uds

    # 9. Construir respuesta · MPs
    def _urgencia_de(deficits_por_h):
        for h in horizontes:
            if deficits_por_h.get(h, 0) > 0.01:
                if h <= 15:
                    return ('CRITICO', h)
                if h <= 30:
                    return ('URGENTE', h)
                if h <= 90:
                    return ('VIGILAR', h)
                return ('PLANIFICAR', h)
        return ('OK', None)

    # FIX 2-jun-2026 audit abastecimiento (P0 · M1) · COLAPSAR la demanda al código
    # de BODEGA resuelto. Antes consumo_mp se llaveaba por el material_id CRUDO de
    # fórmula → la demanda de un mismo material quedaba PARTIDA entre 2-4 códigos (el
    # bridge solo se aplicaba al stock, no a la demanda) → déficit subestimado +
    # pendiente de compra no acreditado (pendientes_mp está por código de bodega) +
    # proveedor vacío + números distintos a auditar-minimos. Resolvemos cada código
    # (id-con-mov → bridge → nombre/alias) y sumamos los horizontes.
    try:
        _consumo_col = {}
        for _cod, _cons in consumo_mp.items():
            _nom = (mp_info.get(_cod, {}) or {}).get('nombre', '') or ''
            _bod = _resolver_material_bodega(c, _cod, _nom) or _cod
            _acc = _consumo_col.setdefault(_bod, {h: 0.0 for h in horizontes})
            for h in horizontes:
                _acc[h] = _acc.get(h, 0.0) + float(_cons.get(h, 0) or 0)
        if _consumo_col:
            consumo_mp = _consumo_col
    except Exception:
        log.warning('colapso consumo_mp por bodega falló · uso crudo', exc_info=True)

    items_out_mp = []
    if incluir_mp:
        for cod, consumo in consumo_mp.items():
            info = mp_info.get(cod, {'nombre': cod, 'proveedor': '',
                                     'lead_time_dias': 14, 'buffer_dias': 30})
            # FIX 1-jun-2026 audit · lookup canónico 5-tier (antes solo id → MP con
            # variante de nombre no bridgeada daba stock 0 → déficit falso).
            stock_g = _lookup_stock_5tier(stock_mp, cod, info.get('nombre') or '')
            pend_g = float(pendientes_mp.get(cod.upper(), 0) or 0)
            disponible = stock_g + pend_g
            deficits = {h: round(max(consumo[h] - disponible, 0), 1) for h in horizontes}
            urg, h_urg = _urgencia_de(deficits)
            if urg == 'OK' and max(consumo.values()) <= 0.01:
                continue  # sin consumo · no mostrar
            items_out_mp.append({
                'codigo': cod,
                'nombre': info['nombre'],
                'nombre_inci': info.get('nombre_inci', ''),
                'proveedor_sugerido': info['proveedor'],
                'tipo': 'MP',
                'stock_actual_g': round(stock_g, 1),
                'pendiente_compras_g': round(pend_g, 1),
                'consumo': {str(h): round(consumo[h], 1) for h in horizontes},
                'deficit': {str(h): deficits[h] for h in horizontes},
                'urgencia': urg,
                'horizonte_quiebre_dias': h_urg,
                'lead_time_dias': info.get('lead_time_dias', 14),
                'buffer_dias': info.get('buffer_dias', 30),
                'solicitudes_en_curso': solicitudes_en_curso.get(cod.upper(), []),
            })

    items_out_mee = []
    if incluir_mee:
        for cod, consumo in consumo_mee.items():
            info = mee_info.get(cod, {'descripcion': cod, 'proveedor': ''})
            stock_u = float(stock_mee.get(cod, 0) or 0)
            pend_u = float(pendientes_mee.get(cod, 0) or 0)
            disponible = stock_u + pend_u
            deficits = {h: round(max(consumo[h] - disponible, 0), 1) for h in horizontes}
            urg, h_urg = _urgencia_de(deficits)
            if urg == 'OK' and max(consumo.values()) <= 0.01:
                continue
            items_out_mee.append({
                'codigo': cod,
                'nombre': info['descripcion'],
                'proveedor_sugerido': info['proveedor'],
                'tipo': 'MEE',
                'stock_actual_u': round(stock_u, 1),
                'pendiente_compras_u': round(pend_u, 1),
                'consumo': {str(h): round(consumo[h], 1) for h in horizontes},
                'deficit': {str(h): deficits[h] for h in horizontes},
                'urgencia': urg,
                'horizonte_quiebre_dias': h_urg,
                'lead_time_dias': 14,  # MEE default · TODO leer de mee_lead_time_config
                'buffer_dias': 30,
                'solicitudes_en_curso': solicitudes_en_curso.get(cod, []),
            })

    # Ordenar: urgencia primero (CRITICO > URGENTE > VIGILAR > PLANIFICAR > OK)
    _orden_urg = {'CRITICO': 0, 'URGENTE': 1, 'VIGILAR': 2, 'PLANIFICAR': 3, 'OK': 4}
    items_out_mp.sort(key=lambda x: (_orden_urg.get(x['urgencia'], 9),
                                     x['horizonte_quiebre_dias'] or 9999))
    items_out_mee.sort(key=lambda x: (_orden_urg.get(x['urgencia'], 9),
                                      x['horizonte_quiebre_dias'] or 9999))

    # Resumen por horizonte
    resumen = {}
    for h in horizontes:
        n_mp_def = sum(1 for it in items_out_mp if it['deficit'][str(h)] > 0.01)
        n_mee_def = sum(1 for it in items_out_mee if it['deficit'][str(h)] > 0.01)
        resumen[str(h)] = {
            'n_mp_con_deficit': n_mp_def,
            'n_mee_con_deficit': n_mee_def,
            'n_total_con_deficit': n_mp_def + n_mee_def,
        }

    # FIX UX 24-may-2026 noche · Sebastián: "no hay solo 24 hay 140 · en
    # promedio tenemos 3 producciones a la semana". Necesita desglose claro
    # por origen + tasa de lotes/semana para no confundir cuántos lotes hay
    # vs cuántas MPs con déficit (las 2 cifras antes se mezclaban).
    n_fijas = sum(1 for r in prod_rows if r[6] in ('eos_plan', 'eos_b2b', 'eos_retroactivo'))
    n_sugeridas = sum(1 for r in prod_rows if r[6] in ('eos_canonico', 'auto_plan', 'sugerido', 'calendar', 'manual'))
    n_otras = len(prod_rows) - n_fijas - n_sugeridas

    # Tasa lotes/semana (en 90d para tener data estable)
    from datetime import date as _d2, timedelta as _td2
    hoy_d = _d2.fromisoformat(hoy_iso) if hoy_iso else _d2.today()
    cutoff_90 = (hoy_d + _td2(days=90)).isoformat()
    lotes_en_90d = sum(1 for r in prod_rows if r[2] and str(r[2])[:10] <= cutoff_90)
    lotes_por_semana_90d = round(lotes_en_90d / (90 / 7), 1) if lotes_en_90d > 0 else 0

    # Fecha del último lote programado para mostrar cobertura real
    fechas_futuras = [str(r[2])[:10] for r in prod_rows if r[2]]
    ultimo_lote_fecha = max(fechas_futuras) if fechas_futuras else None
    cobertura_dias = 0
    if ultimo_lote_fecha:
        try:
            cobertura_dias = (_d2.fromisoformat(ultimo_lote_fecha) - hoy_d).days
        except Exception:
            pass

    return jsonify({
        'hoy': hoy_iso,
        'horizontes': horizontes,
        'modo': modo,
        'n_producciones_fijas': n_fijas,
        'n_producciones_sugeridas': n_sugeridas,
        'n_producciones_otras': n_otras,
        'n_producciones_total': len(prod_rows),
        'n_pedidos_b2b_pendientes': len(b2b_rows),
        'lotes_por_semana_90d': lotes_por_semana_90d,
        'ultimo_lote_fecha': ultimo_lote_fecha,
        'cobertura_dias': cobertura_dias,
        'tipo': tipo,
        'mps': items_out_mp,
        'mees': items_out_mee,
        'resumen_por_horizonte': resumen,
        'productos_sin_lote_size': sorted(productos_sin_lote_size),
        'productos_multi_volumen': sorted(productos_multi_volumen),
        # FIX 24-may-2026 noche · diagnóstico de match producto↔fórmula
        # · si muchos lotes están en sin_formula, el cálculo subestima
        # gravemente. Util para que admin vea cuántos lotes el sistema
        # NO pudo cruzar con formula_items (por nombre que no matchea).
        'lotes_con_formula': matched_lotes,
        'lotes_sin_formula': len(sin_formula_lotes),
        'productos_sin_match_formula': sorted(set(sin_formula_lotes))[:30],
    })


@bp.route('/api/abastecimiento/trail-mp/<codigo_mp>', methods=['GET'])
def abastecimiento_trail_mp(codigo_mp):
    """FIX UX 24-may-2026 noche · Sebastián: 'no muestra la realidad ·
    pensaria yo que debe tomar producto por producto enlazar formula
    e ir sumando'.

    Trail completo de UNA materia prima: devuelve la lista de productos
    que la usan + cuántos lotes programados hay de cada uno + gramos
    calculados de cada lote = total. Así se ve POR QUÉ sale X gramos
    a 365d (e.g., 'Centella solo tiene 1 lote HYDRAPEPTIDE futuro,
    AZH y SUERO NIACINAMIDA están en BD pero sin lotes programados').

    Endpoint solo lectura. Acceso compras_user."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    c = conn.cursor()
    codigo_up = (codigo_mp or '').strip().upper()

    # 1. Info de la MP
    mp_info = c.execute(
        """SELECT codigo_mp,
                  COALESCE(nombre_comercial, nombre_inci, codigo_mp),
                  COALESCE(proveedor, ''),
                  COALESCE(nombre_inci, '')
           FROM maestro_mps
           WHERE UPPER(TRIM(codigo_mp)) = ?""",
        (codigo_up,),
    ).fetchone()
    if not mp_info:
        return jsonify({'error': f'MP {codigo_mp} no existe'}), 404

    # 2. Productos que usan esta MP (de formula_items)
    productos_que_usan = c.execute(
        """SELECT producto_nombre,
                  COALESCE(porcentaje, 0),
                  COALESCE(cantidad_g_por_lote, 0)
           FROM formula_items
           WHERE UPPER(TRIM(material_id)) = ?
           ORDER BY producto_nombre""",
        (codigo_up,),
    ).fetchall()

    # 3. Para cada producto, contar lotes futuros + lote_size
    from datetime import date as _d, timedelta as _td
    hoy = _d.today().isoformat()
    cutoff = (_d.today() + _td(days=365)).isoformat()
    productos_trail = []
    total_gramos_365 = 0.0
    for prod_nom, pct, g_por_lote in productos_que_usan:
        # lote_size del producto
        lote_size_row = c.execute(
            "SELECT COALESCE(lote_size_kg, 0) FROM formula_headers "
            "WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))",
            (prod_nom,),
        ).fetchone()
        lote_size_kg = float(lote_size_row[0]) if lote_size_row else 0
        # Lotes futuros programados (Fijas + Sugeridas + B2B · sin pasados ni cancelados)
        # FIX 24-may noche · normalizar espacios en Python (SQL no colapsa
        # espacios dobles interiores). Bringo todos los lotes del horizonte
        # y filtro por nombre normalizado.
        def _np(s):
            return ' '.join((s or '').strip().upper().split())
        prod_norm_target = _np(prod_nom)
        lotes_rows_raw = c.execute(
            """SELECT id, fecha_programada, COALESCE(cantidad_kg, 0),
                      COALESCE(estado, ''), COALESCE(origen, ''), producto
               FROM produccion_programada
               WHERE LOWER(COALESCE(estado,'')) NOT IN
                     ('cancelado','completado','esperando_recurso')
                 AND COALESCE(inventario_descontado_at,'') = ''
                 AND fecha_programada >= ? AND fecha_programada <= ?
               ORDER BY fecha_programada ASC""",
            (hoy, cutoff),
        ).fetchall()
        lotes_rows = [(r[0], r[1], r[2], r[3], r[4]) for r in lotes_rows_raw
                       if _np(r[5]) == prod_norm_target]
        lotes_detalle = []
        gramos_producto_total = 0.0
        for lid, lfecha, lkg, lest, lorigen in lotes_rows:
            lkg_f = float(lkg or 0)
            # Calcular gramos como en el endpoint principal
            # FIX 24-may noche · priorizar porcentaje (g_por_lote seed roto · 674 items)
            if pct > 0:
                gramos = (float(pct) / 100.0) * lkg_f * 1000.0
            elif g_por_lote > 0 and lote_size_kg > 0:
                gramos = float(g_por_lote) * (lkg_f / lote_size_kg)
            else:
                gramos = 0
            gramos_producto_total += gramos
            lotes_detalle.append({
                'lote_id': lid,
                'fecha': (lfecha or '')[:10],
                'cantidad_kg': lkg_f,
                'estado': lest,
                'origen': lorigen,
                'gramos': round(gramos, 2),
            })
        total_gramos_365 += gramos_producto_total
        productos_trail.append({
            'producto_nombre': prod_nom,
            'porcentaje': pct,
            'cantidad_g_por_lote': g_por_lote,
            'lote_size_kg': lote_size_kg,
            'n_lotes_futuros': len(lotes_rows),
            'gramos_total_365d': round(gramos_producto_total, 2),
            'lotes_detalle': lotes_detalle[:30],
            'mas_lotes': max(0, len(lotes_rows) - 30),
        })

    # 4. Stock canónico actual + pendientes (igual que endpoint principal)
    stock_mp = _get_mp_stock(conn)
    stock_actual = float(stock_mp.get(codigo_up, 0))
    pendiente = 0.0
    try:
        for r in c.execute(
            """SELECT COALESCE(SUM(sci.cantidad_solicitada_g), 0)
               FROM solicitudes_compra_items sci
               JOIN solicitudes_compra sc ON sc.id = sci.solicitud_id
               WHERE UPPER(TRIM(sci.codigo_mp)) = ?
                 AND sc.estado IN ('Pendiente', 'Aprobada')
                 AND COALESCE(sc.numero_oc, '') = ''""",
            (codigo_up,),
        ).fetchall():
            pendiente += float(r[0] or 0)
    except Exception:
        pass
    try:
        for r in c.execute(
            """SELECT COALESCE(SUM(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g, 0)), 0)
               FROM ordenes_compra_items oci
               JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
               WHERE UPPER(TRIM(oci.codigo_mp)) = ?
                 AND oc.estado IN ('Borrador', 'Revisada', 'Autorizada', 'Parcial')""",
            (codigo_up,),
        ).fetchall():
            pendiente += float(r[0] or 0)
    except Exception:
        pass

    productos_trail.sort(key=lambda x: -x['gramos_total_365d'])
    return jsonify({
        'codigo_mp': codigo_up,
        'nombre_comercial': mp_info[1],
        'nombre_inci': mp_info[3],
        'proveedor': mp_info[2],
        'stock_actual_g': round(stock_actual, 1),
        'pendiente_compras_g': round(pendiente, 1),
        'total_consumo_365d_g': round(total_gramos_365, 2),
        'deficit_365d_g': round(max(0, total_gramos_365 - stock_actual - pendiente), 2),
        'n_productos_que_usan': len(productos_trail),
        'productos': productos_trail,
    })


@bp.route('/api/abastecimiento/export-excel', methods=['GET'])
def abastecimiento_export_excel():
    """Descarga Excel con el consumo de abastecimiento para enviar a Alejandro.

    Sebastián 23-may-2026: 'poder descargar un excel o pdf · Alejandro a
    veces dice manden qué necesitan · que salga consolidado para enviarle'.

    Mismos query params que consumo-horizontes (modo, tipo, horizontes).
    Filtros UI adicionales (búsqueda, urgencia, proveedor) se pasan también
    para que el Excel coincida con lo que ve Sebastián en pantalla.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl no disponible'}), 500
    from flask import current_app, send_file
    import io as _io
    from datetime import datetime as _dt

    # Llamar internamente al endpoint consumo-horizontes (misma lógica)
    qs = []
    for k in ('modo', 'tipo', 'horizontes'):
        v = request.args.get(k)
        if v:
            qs.append(f'{k}={v}')
    url_interno = '/api/abastecimiento/consumo-horizontes'
    if qs:
        url_interno += '?' + '&'.join(qs)
    with current_app.test_request_context(url_interno):
        # Copiar la sesión del request actual para que el endpoint pase auth
        from flask import session as _sess
        _sess['compras_user'] = session.get('compras_user', 'sistema')
        resp = abastecimiento_consumo_horizontes()
    if isinstance(resp, tuple):
        return resp  # error
    d = resp.get_json() or {}

    # Aplicar filtros del cliente (UI)
    f_busq = (request.args.get('busqueda', '') or '').lower().strip()
    f_urg = request.args.get('urgencia', 'TODAS') or 'TODAS'
    f_prov = request.args.get('proveedor', 'TODOS') or 'TODOS'
    f_tipo = request.args.get('tipo_filtro', 'TODOS') or 'TODOS'
    def _pasa_filtros(it):
        if f_urg != 'TODAS' and it['urgencia'] != f_urg:
            return False
        prov = it.get('proveedor_sugerido') or '(sin proveedor)'
        if f_prov != 'TODOS' and prov != f_prov:
            return False
        if f_tipo != 'TODOS' and it['tipo'] != f_tipo:
            return False
        if f_busq:
            blob = (it['codigo'] + ' ' + (it.get('nombre') or '') + ' ' +
                    (it.get('proveedor_sugerido') or '')).lower()
            if f_busq not in blob:
                return False
        return True

    items_mp = [it for it in (d.get('mps') or []) if _pasa_filtros(it)]
    items_mee = [it for it in (d.get('mees') or []) if _pasa_filtros(it)]
    horizontes = d.get('horizontes') or [15, 30, 60, 90, 120, 180, 365]

    # Construir Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='7C3AED')
    title_fill = PatternFill('solid', fgColor='5B21B6')
    crit_fill = PatternFill('solid', fgColor='FEE2E2')
    urg_fill = PatternFill('solid', fgColor='FFF7ED')
    vig_fill = PatternFill('solid', fgColor='FEFCE8')
    ok_fill = PatternFill('solid', fgColor='F0FDF4')

    def _fill_for_urg(u):
        return {
            'CRITICO': crit_fill, 'URGENTE': urg_fill,
            'VIGILAR': vig_fill, 'PLANIFICAR': ok_fill, 'OK': ok_fill,
        }.get(u, ok_fill)

    # SHEET 1 · Resumen completo
    ws = wb.create_sheet('Abastecimiento')
    fecha_str = _dt.now().strftime('%Y-%m-%d %H:%M')
    modo_str = d.get('modo', 'comprometido')
    n_cols = 9 + len(horizontes) + 1  # FIX 23-may · +1 por columna INCI
    # Título
    ws.cell(row=1, column=1, value=f'Abastecimiento · Modo {modo_str.upper()} · {fecha_str}')
    ws.cell(row=1, column=1).font = Font(size=14, bold=True, color='FFFFFF')
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 24
    # Subtítulo
    ws.cell(row=2, column=1, value=f'{d.get("n_producciones_fijas",0)} producciones Fijas · {d.get("n_pedidos_b2b_pendientes",0)} pedidos B2B pendientes · {len(items_mp)} MPs + {len(items_mee)} MEEs con consumo')
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='6B7280')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    # Encabezados (fila 4)
    # FIX 23-may-2026 · Sebastián pidió columna INCI (nombre químico
    # estándar · útil para verificar fórmula y para compras a proveedores
    # internacionales que solo conocen el INCI)
    hdr = ['Código', 'Nombre', 'INCI', 'Tipo', 'Proveedor', 'LT (días)',
           'Stock actual', 'En cola', 'Urgencia']
    # AUDIT FIX 23-may · header dinámico · si 60d no está en horizontes
    # busca el del medio (no mentir al usuario)
    target_cubrir = 60 if 60 in horizontes else horizontes[len(horizontes) // 2]
    for h in horizontes:
        hdr.append(f'Déficit {h}d')
    hdr.append(f'Cant. a pedir (cubrir {target_cubrir}d)')
    for col, val in enumerate(hdr, start=1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[4].height = 30

    # Filas
    todos = items_mp + items_mee
    todos.sort(key=lambda x: ({'CRITICO':0,'URGENTE':1,'VIGILAR':2,'PLANIFICAR':3,'OK':4}.get(x['urgencia'], 9),
                              x.get('horizonte_quiebre_dias') or 9999))
    r = 5
    for it in todos:
        urg = it['urgencia']
        fill = _fill_for_urg(urg)
        stock_key = 'stock_actual_g' if it['tipo']=='MP' else 'stock_actual_u'
        cola_key = 'pendiente_compras_g' if it['tipo']=='MP' else 'pendiente_compras_u'
        unit = 'g' if it['tipo']=='MP' else 'u'
        # Sugerencia de cantidad · déficit del horizonte target_cubrir
        # AUDIT FIX 23-may · float() defensivo · evita None<=0 TypeError
        # y usa target_cubrir (puede no ser 60) en lugar de '60' hardcoded
        deficit_dict = it.get('deficit', {})
        cant_sug = float(deficit_dict.get(str(target_cubrir)) or 0)
        if cant_sug <= 0:
            for h in horizontes:
                v = float(deficit_dict.get(str(h)) or 0)
                if v > 0:
                    cant_sug = v
                    break
        valores = [
            it['codigo'], it.get('nombre',''),
            it.get('nombre_inci','') or '—',
            it['tipo'],
            it.get('proveedor_sugerido','') or '(sin proveedor)',
            it.get('lead_time_dias', 14),
            it.get(stock_key, 0), it.get(cola_key, 0), urg,
        ]
        for h in horizontes:
            valores.append(float(deficit_dict.get(str(h)) or 0))
        valores.append(round(cant_sug, 1))
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            # Col 9 = Urgencia (antes era 8 · INCI shifted everything +1)
            if col == 9:  # urgencia
                cell.fill = fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            elif col in (1, 4):  # código/tipo (tipo ahora col 4 por INCI)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Consolas')
            elif col == 3:  # INCI
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(italic=True, size=10, color='6B7280')
            elif col >= 6:  # números (era >=5 · ahora 6 por INCI)
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.0'
        r += 1

    # Anchos
    widths = [12, 32, 28, 6, 22, 9, 13, 11, 12]
    for h in horizontes:
        widths.append(12)
    widths.append(15)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A5'

    # SHEET 2 · Agrupado por proveedor (lo que Catalina/Alejandro envían)
    ws2 = wb.create_sheet('Por proveedor')
    ws2.cell(row=1, column=1, value='Consolidado por proveedor (lo que hay que comprar)')
    ws2.cell(row=1, column=1).font = Font(size=14, bold=True, color='FFFFFF')
    ws2.cell(row=1, column=1).fill = title_fill
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws2.row_dimensions[1].height = 24
    hdr2 = ['Proveedor', 'Código', 'Nombre', 'Tipo', 'Cant. a pedir', 'Urgencia']
    for col, val in enumerate(hdr2, start=1):
        cell = ws2.cell(row=3, column=col, value=val)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Agrupar por proveedor con cant_sug
    grupos = {}
    for it in todos:
        cant_sug = it.get('deficit', {}).get('60', 0)
        if cant_sug <= 0:
            for h in horizontes:
                v = it['deficit'].get(str(h), 0)
                if v > 0:
                    cant_sug = v
                    break
        if cant_sug <= 0:
            continue
        prov = it.get('proveedor_sugerido') or '(sin proveedor)'
        grupos.setdefault(prov, []).append((it, cant_sug))

    r = 4
    for prov in sorted(grupos.keys()):
        for it, cant in grupos[prov]:
            urg = it['urgencia']
            valores = [prov, it['codigo'], it.get('nombre',''), it['tipo'], round(cant, 1), urg]
            for col, val in enumerate(valores, start=1):
                cell = ws2.cell(row=r, column=col, value=val)
                cell.border = border
                if col == 6:
                    cell.fill = _fill_for_urg(urg)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                elif col == 5:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0.0'
                elif col == 2:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(name='Consolas')
            r += 1
        # Fila vacía entre proveedores
        r += 1

    for col, w in enumerate([24, 12, 38, 6, 14, 12], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A4'

    # Stream out
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f'abastecimiento_{modo_str}_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/abastecimiento/consumo-bruto-excel', methods=['GET'])
def abastecimiento_consumo_bruto_excel():
    """Excel para Alejandro · CONSUMO BRUTO total de MP/MEE por horizonte.

    Diferente al export-excel normal (que muestra déficit a comprar
    restando stock + pendiente): aquí se muestra el consumo TOTAL en
    gramos (MP) o unidades (MEE) que se va a consumir según producciones
    programadas, SIN restar inventario.

    Sebastián 23-may-2026: 'Alejandro quiere el gasto total en gramos
    de las materias primas para el saber según los horizontes · sin
    contar lo que tiene el inventario'.

    Mismo modelo que el export-excel: respeta filtros UI + modo dual
    (comprometido/run_rate).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl no disponible'}), 500
    from flask import current_app, send_file
    import io as _io
    from datetime import datetime as _dt

    # Llamar internamente al endpoint consumo-horizontes (misma lógica)
    qs = []
    for k in ('modo', 'tipo', 'horizontes'):
        v = request.args.get(k)
        if v:
            qs.append(f'{k}={v}')
    url_interno = '/api/abastecimiento/consumo-horizontes'
    if qs:
        url_interno += '?' + '&'.join(qs)
    with current_app.test_request_context(url_interno):
        from flask import session as _sess
        _sess['compras_user'] = session.get('compras_user', 'sistema')
        resp = abastecimiento_consumo_horizontes()
    if isinstance(resp, tuple):
        return resp
    d = resp.get_json() or {}
    horizontes = d.get('horizontes') or [15, 30, 60, 90, 120, 180, 365]
    items_mp = d.get('mps') or []
    items_mee = d.get('mees') or []

    # Aplicar filtros UI (mismo patrón que export-excel normal)
    f_busq = (request.args.get('busqueda', '') or '').lower().strip()
    f_urg = request.args.get('urgencia', 'TODAS') or 'TODAS'
    f_prov = request.args.get('proveedor', 'TODOS') or 'TODOS'
    f_tipo = request.args.get('tipo_filtro', 'TODOS') or 'TODOS'
    def _pasa(it):
        # AUDIT FIX 23-may · .get() defensivo · KeyError potencial
        if f_urg != 'TODAS' and it.get('urgencia', 'OK') != f_urg:
            return False
        prov = it.get('proveedor_sugerido') or '(sin proveedor)'
        if f_prov != 'TODOS' and prov != f_prov:
            return False
        if f_tipo != 'TODOS' and it.get('tipo', '') != f_tipo:
            return False
        if f_busq:
            blob = (str(it.get('codigo', '')) + ' ' + (it.get('nombre') or '') + ' ' +
                    (it.get('proveedor_sugerido') or '')).lower()
            if f_busq not in blob:
                return False
        return True
    items_mp = [it for it in items_mp if _pasa(it)]
    items_mee = [it for it in items_mee if _pasa(it)]

    # Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='065F46')
    title_fill = PatternFill('solid', fgColor='047857')
    total_fill = PatternFill('solid', fgColor='D1FAE5')
    mp_color = '0891B2'
    mee_color = '7C3AED'

    # AUDIT FIX 23-may · cursor para query de fecha última Fija
    conn = get_db(); c = conn.cursor()

    # SHEET 1 · Consumo bruto
    ws = wb.create_sheet('Consumo bruto')
    fecha_str = _dt.now().strftime('%Y-%m-%d %H:%M')
    modo = d.get('modo', 'comprometido')
    n_cols = 5 + len(horizontes)
    # FIX 24-may noche · título explícito: SIN descontar inventario
    ws.cell(row=1, column=1, value=f'Consumo TOTAL del Calendario · SIN descontar inventario · Modo {modo.upper()} · {fecha_str}')
    ws.cell(row=1, column=1).font = Font(size=14, bold=True, color='FFFFFF')
    ws.cell(row=1, column=1).fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 26

    # Fecha del último lote programado · para que el usuario sepa hasta
    # dónde llega el calendario realmente
    fecha_ultimo_lote = None
    try:
        r_ult = c.execute(
            "SELECT MAX(fecha_programada) FROM produccion_programada "
            "WHERE LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado','esperando_recurso') "
            "AND COALESCE(inventario_descontado_at,'') = ''"
        ).fetchone()
        if r_ult and r_ult[0]:
            fecha_ultimo_lote = str(r_ult[0])[:10]
    except Exception:
        pass

    # Desglose lotes por origen (alineado con UI nueva)
    n_fijas = d.get('n_producciones_fijas', 0)
    n_sugeridas = d.get('n_producciones_sugeridas', 0)
    n_b2b = d.get('n_pedidos_b2b_pendientes', 0)
    n_total = d.get('n_producciones_total', n_fijas + n_sugeridas)
    lotes_sem = d.get('lotes_por_semana_90d', 0)
    cobertura = d.get('cobertura_dias', 0)

    subtitulo = (
        f'{n_total} lotes totales · {n_fijas} Fijas + {n_sugeridas} Sugeridas + {n_b2b} B2B pendientes · '
        f'{lotes_sem} lotes/sem (90d) · cobertura {cobertura}d · '
        f'último lote {fecha_ultimo_lote or "—"} · '
        f'consumo acumulativo desde hoy · MP en gramos · MEE en unidades · '
        f'NO se resta stock actual ni órdenes en curso'
    )
    ws.cell(row=2, column=1, value=subtitulo)
    ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='6B7280')
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 32

    # Aviso si modo comprometido sin actividad lejana
    aviso_modo = ''
    if modo == 'comprometido' and fecha_ultimo_lote and cobertura < 180:
        aviso_modo = (f'⚠ Cobertura {cobertura}d · horizontes >cobertura muestran '
                       'el mismo total porque no hay más lotes programados · '
                       'usá Run-rate o llená el calendario')
        ws.cell(row=3, column=1, value=aviso_modo)
        ws.cell(row=3, column=1).font = Font(size=10, italic=True, color='B45309', bold=True)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)

    # Encabezados fila 4
    # FIX 23-may-2026 · Sebastián pidió columna INCI · útil para Alejandro
    # cuando contacta proveedores internacionales que usan INCI estándar
    # + Proveedor (sugerencia auditoría · Alejandro contacta proveedores)
    hdr = ['Código', 'Nombre', 'INCI', 'Proveedor', 'Tipo']
    for h in horizontes:
        hdr.append(f'{h}d')
    for col, val in enumerate(hdr, start=1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[4].height = 26

    # Filas
    todos = []
    for it in items_mp:
        todos.append({
            'codigo': it['codigo'],
            'nombre': it.get('nombre', ''),
            'nombre_inci': it.get('nombre_inci', ''),
            'proveedor': it.get('proveedor_sugerido', '') or '',
            'tipo': 'MP',
            'unit': 'g',
            'consumo': it.get('consumo', {}),
        })
    for it in items_mee:
        todos.append({
            'codigo': it['codigo'],
            'nombre': it.get('nombre', ''),
            'nombre_inci': '',  # MEE no tiene INCI
            'proveedor': it.get('proveedor_sugerido', '') or '',
            'tipo': 'MEE',
            'unit': 'u',
            'consumo': it.get('consumo', {}),
        })
    # Ordenar por consumo del horizonte más largo desc (ver impacto anual)
    horz_max = max(horizontes)
    todos.sort(key=lambda x: -float(x['consumo'].get(str(horz_max), 0) or 0))

    # AUDIT FIX 23-may · mensaje cuando no hay items
    if not todos:
        ws.cell(row=5, column=1, value='Sin consumo programado en el horizonte').font = Font(italic=True, color='6B7280')

    r = 5
    totales_h_mp = {h: 0.0 for h in horizontes}
    totales_h_mee = {h: 0.0 for h in horizontes}
    for it in todos:
        valores = [it['codigo'], it['nombre'],
                   it['nombre_inci'] or '—',
                   it['proveedor'] or '(sin proveedor)',
                   it['tipo']]
        for h in horizontes:
            v = float(it['consumo'].get(str(h), 0) or 0)
            valores.append(v)
            if it['tipo'] == 'MP':
                totales_h_mp[h] += v
            else:
                totales_h_mee[h] += v
        for col, val in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            if col == 1:
                cell.font = Font(name='Consolas')
                cell.alignment = Alignment(horizontal='center')
            elif col == 3:  # INCI
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(italic=True, size=10, color='6B7280')
            elif col == 4:  # Proveedor
                cell.alignment = Alignment(horizontal='left')
            elif col == 5:  # Tipo
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True, color=mp_color if val == 'MP' else mee_color)
            elif col >= 6:  # números
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
        r += 1

    # Filas TOTALES · AUDIT FIX · borders en todas las celdas
    def _tot_celda_vacia(row, col):
        cell = ws.cell(row=row, column=col, value='')
        cell.fill = total_fill
        cell.border = border
    ws.cell(row=r, column=1, value='TOTAL MP (g)').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=mp_color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=1).border = border
    for col in (2, 3, 4, 5):
        _tot_celda_vacia(r, col)
    for i, h in enumerate(horizontes):
        cell = ws.cell(row=r, column=6 + i, value=totales_h_mp[h])
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
    r += 1
    ws.cell(row=r, column=1, value='TOTAL MEE (u)').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=mee_color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=1).border = border
    for col in (2, 3, 4, 5):
        _tot_celda_vacia(r, col)
    for i, h in enumerate(horizontes):
        cell = ws.cell(row=r, column=6 + i, value=totales_h_mee[h])
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border

    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 28  # INCI
    ws.column_dimensions['D'].width = 22  # Proveedor
    ws.column_dimensions['E'].width = 6   # Tipo
    for i in range(len(horizontes)):
        ws.column_dimensions[get_column_letter(6 + i)].width = 14
    ws.freeze_panes = 'F5'

    # AUDIT FIX · n_cols correcto para merge_cells
    # (5 cols base + horizontes · ajustar arriba si necesario)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f'consumo_bruto_alejandro_{modo}_{_dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/abastecimiento/solicitar-items', methods=['POST'])
def abastecimiento_solicitar_items():
    """Crea SOLs agrupadas por proveedor a partir de items seleccionados
    en el tab Abastecimiento.

    Sebastián 23-may-2026: 'centro de solicitudes a compras' · Sebastián
    elige items y cantidades en la UI · este endpoint crea SOLs.

    Body:
      items: [{tipo: 'mp'|'mee', codigo, cantidad, proveedor_sugerido?}]
      agrupar_por_proveedor: bool (default true · 1 SOL por proveedor)
      urgencia: 'Alta'|'Normal'|'Baja' (default 'Normal')
      cubrir_dias: int (info en obs · no afecta cantidad)

    Reusa el patrón de solicitar_faltantes_bulk (audit_log + numero_unique
    con retry) pero recibe items pre-seleccionados con cantidades.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', 'sistema')
    d = request.get_json(silent=True) or {}
    items_in = d.get('items') or []
    if not isinstance(items_in, list) or not items_in:
        return jsonify({'error': 'items[] requerido'}), 400
    agrupar = bool(d.get('agrupar_por_proveedor', True))
    urgencia_norm = (d.get('urgencia') or 'Normal').strip()
    if urgencia_norm not in ('Alta', 'Normal', 'Baja'):
        urgencia_norm = 'Normal'
    cubrir_dias = d.get('cubrir_dias')
    try:
        cubrir_dias = int(cubrir_dias) if cubrir_dias is not None else None
    except Exception:
        cubrir_dias = None

    # Validar items + agrupar
    grupos = {}  # proveedor → {'mps': [...], 'mees': [...]}
    conn = get_db(); c = conn.cursor()
    # Cargar info MP/MEE para nombres y proveedor sugerido si falta
    mp_info_map = {}
    try:
        for r in c.execute("""
            SELECT mm.codigo_mp,
                   COALESCE(mm.nombre_comercial, mm.nombre_inci, mm.codigo_mp),
                   COALESCE(NULLIF(TRIM(mlt.proveedor_principal),''),
                            mm.proveedor, '')
            FROM maestro_mps mm
            LEFT JOIN mp_lead_time_config mlt ON mlt.material_id = mm.codigo_mp
        """).fetchall():
            mp_info_map[str(r[0]).strip().upper()] = {
                'nombre': r[1] or r[0],
                'proveedor': (r[2] or '').strip(),
            }
    except sqlite3.OperationalError:
        pass
    mee_info_map = {}
    try:
        for r in c.execute("""
            SELECT codigo, COALESCE(descripcion,''), COALESCE(proveedor,'')
            FROM maestro_mee
        """).fetchall():
            mee_info_map[str(r[0]).strip().upper()] = {
                'nombre': r[1] or r[0],
                'proveedor': (r[2] or '').strip(),
            }
    except sqlite3.OperationalError:
        pass

    for it in items_in:
        tipo = (it.get('tipo') or '').lower()
        codigo = str(it.get('codigo') or '').strip().upper()
        try:
            cantidad = float(it.get('cantidad') or 0)
        except Exception:
            cantidad = 0
        if not codigo or cantidad <= 0 or tipo not in ('mp', 'mee'):
            continue
        prov_override = (it.get('proveedor_sugerido') or '').strip()
        if tipo == 'mp':
            info = mp_info_map.get(codigo, {'nombre': codigo, 'proveedor': ''})
        else:
            info = mee_info_map.get(codigo, {'nombre': codigo, 'proveedor': ''})
        proveedor = prov_override or info['proveedor'] or '__SIN_PROVEEDOR__'
        grupo_key = proveedor if agrupar else f"{proveedor}::{codigo}"
        g_node = grupos.setdefault(grupo_key, {'proveedor': proveedor,
                                                'mps': [], 'mees': []})
        if tipo == 'mp':
            g_node['mps'].append({
                'codigo_mp': codigo,
                'nombre': info['nombre'],
                'cantidad_g': cantidad,
            })
        else:
            g_node['mees'].append({
                'codigo': codigo,
                'descripcion': info['nombre'],
                'cantidad_u': cantidad,
            })

    if not grupos:
        return jsonify({'error': 'No hay items válidos'}), 400

    # Crear SOLs · patrón retry del solicitar_faltantes_bulk
    from datetime import datetime as _dt
    fecha_now = _dt.now().isoformat()
    year = _dt.now().strftime('%Y')
    obs_base = f"Auto-generada Abastecimiento"
    if cubrir_dias:
        obs_base += f" · cubrir {cubrir_dias}d"

    creadas = []
    try:
        for grupo_key, gnode in sorted(grupos.items()):
            mps_lst = gnode['mps']
            mees_lst = gnode['mees']
            if not mps_lst and not mees_lst:
                continue
            prov = gnode['proveedor']
            prov_label = prov if prov != '__SIN_PROVEEDOR__' else ''
            obs = obs_base
            if prov_label:
                obs += f" · proveedor: {prov_label}"
            obs += f" · {len(mps_lst)} MPs · {len(mees_lst)} MEEs"
            categoria = 'Materia Prima'
            if not mps_lst and mees_lst:
                categoria = 'Empaque'

            # Retry numero único
            numero = None
            ultimo_err = None
            for retry in range(6):
                c.execute(
                    "SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) "
                    "FROM solicitudes_compra WHERE numero LIKE ?",
                    (f"SOL-{year}-%",),
                )
                num = (c.fetchone()[0] or 0) + 1 + retry
                candidato = f"SOL-{year}-{num:04d}"
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra
                          (numero, fecha, estado, solicitante, urgencia, observaciones,
                           area, empresa, categoria, tipo)
                        VALUES (?, ?, 'Pendiente', ?, ?, ?, 'Producción', 'Espagiria', ?, 'Compra')
                    """, (candidato, fecha_now, user, urgencia_norm, obs, categoria))
                    numero = candidato
                    break
                except sqlite3.IntegrityError as e:
                    ultimo_err = e
                    continue
            if numero is None:
                raise RuntimeError(f"No se pudo asignar numero SOL · {ultimo_err}")

            items_count = 0
            total_g_mps = 0.0
            for mp in mps_lst:
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado, proveedor_sugerido)
                        VALUES (?, ?, ?, ?, 'g', ?, 0, ?)
                    """, (numero, mp['codigo_mp'], mp['nombre'],
                          mp['cantidad_g'],
                          f"Abastecimiento · {obs_base}",
                          prov_label))
                except sqlite3.OperationalError:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado)
                        VALUES (?, ?, ?, ?, 'g', ?, 0)
                    """, (numero, mp['codigo_mp'], mp['nombre'],
                          mp['cantidad_g'],
                          f"Abastecimiento · {obs_base}"))
                items_count += 1
                total_g_mps += mp['cantidad_g']
            for me in mees_lst:
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado, proveedor_sugerido)
                        VALUES (?, ?, ?, ?, 'unidades', ?, 0, ?)
                    """, (numero, me['codigo'], me['descripcion'],
                          me['cantidad_u'],
                          f"Abastecimiento · {obs_base}",
                          prov_label))
                except sqlite3.OperationalError:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado)
                        VALUES (?, ?, ?, ?, 'unidades', ?, 0)
                    """, (numero, me['codigo'], me['descripcion'],
                          me['cantidad_u'],
                          f"Abastecimiento · {obs_base}"))
                items_count += 1

            try:
                audit_log(
                    c, usuario=user, accion='SOLICITAR_ABASTECIMIENTO',
                    tabla='solicitudes_compra', registro_id=numero,
                    despues={
                        'proveedor': prov_label or '(sin proveedor)',
                        'cubrir_dias': cubrir_dias,
                        'urgencia': urgencia_norm,
                        'mps_count': len(mps_lst),
                        'mees_count': len(mees_lst),
                        'total_g_mps': round(total_g_mps, 1),
                        'agrupar_por_proveedor': agrupar,
                    },
                )
            except Exception:
                pass

            creadas.append({
                'numero': numero,
                'proveedor': prov_label or '(sin proveedor)',
                'mps': len(mps_lst),
                'mees': len(mees_lst),
                'total_items': items_count,
            })
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Error creando SOLs: {e}'}), 500

    return jsonify({
        'ok': True,
        'creadas': creadas,
        'n_sols': len(creadas),
        'mensaje': f'✓ {len(creadas)} SOL(s) creada(s)',
    })


@bp.route('/api/programacion/solicitar-faltantes-bulk', methods=['POST'])
def solicitar_faltantes_bulk():
    """Crea solicitudes_compra agrupadas por proveedor desde los faltantes
    detectados en /api/programacion/producciones-faltantes.

    Body:
      dias: int (default 60) · mismo horizonte usado para detectar faltantes
      urgencia: 'Alta' | 'Normal' | 'Baja' (default 'Alta')
      observaciones_extra: str (opcional)

    Returns:
      201 {ok, solicitudes_creadas: [{numero, proveedor, items_count, total_g}],
           total_proveedores, total_solicitudes}
      400 si no hay faltantes
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in (set(COMPRAS_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Compras/Admin pueden generar solicitudes'}), 403

    body = request.get_json(silent=True) or {}
    try:
        dias = max(7, min(int(body.get('dias', 60)), 365))
    except (ValueError, TypeError):
        dias = 60
    urgencia = (body.get('urgencia') or 'Alta').strip()
    if urgencia not in ('Alta', 'Normal', 'Baja'):
        urgencia = 'Alta'
    obs_extra = (body.get('observaciones_extra') or '').strip()

    # Reutilizar el endpoint anterior llamandolo internamente
    from flask import current_app
    with current_app.test_request_context(
        f'/api/programacion/producciones-faltantes?dias={dias}'
    ):
        # Bypass auth check del helper interno
        from flask import session as _s
        _s['compras_user'] = user
        resp = producciones_faltantes()
    data = resp.get_json() if hasattr(resp, 'get_json') else (resp[0].get_json() if isinstance(resp, tuple) else {})
    faltantes_mps = data.get('faltantes_mps') or []
    faltantes_mees = data.get('faltantes_mees') or []
    if not faltantes_mps and not faltantes_mees:
        return jsonify({
            'ok': True,
            'mensaje': 'No hay faltantes en el horizonte · nada que solicitar',
            'solicitudes_creadas': [],
            'total_proveedores': 0,
            'total_solicitudes': 0,
        }), 200

    # Agrupar por proveedor (sin proveedor sugerido = grupo aparte)
    grupos = {}
    for f in faltantes_mps:
        prov = (f.get('proveedor_sugerido') or '').strip() or '__SIN_PROVEEDOR__'
        grupos.setdefault(prov, {'mps': [], 'mees': []})
        grupos[prov]['mps'].append(f)
    for f in faltantes_mees:
        prov = (f.get('proveedor_sugerido') or '').strip() or '__SIN_PROVEEDOR__'
        grupos.setdefault(prov, {'mps': [], 'mees': []})
        grupos[prov]['mees'].append(f)

    conn = get_db(); c = conn.cursor()
    creadas = []
    fecha_now = datetime.now().isoformat()
    obs_base_horizonte = f"Auto-generada Centro Programación · horizonte {dias}d"
    if obs_extra:
        obs_base_horizonte = obs_extra + ' · ' + obs_base_horizonte

    try:
        for prov, items in sorted(grupos.items()):
            mps_lst = items.get('mps', [])
            mees_lst = items.get('mees', [])
            if not mps_lst and not mees_lst:
                continue
            prov_label = prov if prov != '__SIN_PROVEEDOR__' else ''
            obs = obs_base_horizonte
            if prov_label:
                obs += f" · proveedor: {prov_label}"
            obs += (f" · {len(mps_lst)} MPs · {len(mees_lst)} MEEs")

            # Categoria · si solo hay MEEs y nada de MP, marcamos categoria='Empaque'
            categoria = 'Materia Prima'
            if not mps_lst and mees_lst:
                categoria = 'Empaque'

            # FIX 23-may-2026 · auditoría · race condition · SELECT MAX+1 sin
            # lock chocaba con 2 admins concurrentes (numero UNIQUE mig 84) ·
            # ahora retry loop con IntegrityError detection · idempotente
            year = datetime.now().strftime('%Y')
            numero = None
            ultimo_err = None
            for retry in range(6):
                c.execute(
                    "SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) "
                    "FROM solicitudes_compra WHERE numero LIKE ?",
                    (f"SOL-{year}-%",),
                )
                num = (c.fetchone()[0] or 0) + 1 + retry
                candidato = f"SOL-{year}-{num:04d}"
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra
                          (numero, fecha, estado, solicitante, urgencia, observaciones,
                           area, empresa, categoria, tipo)
                        VALUES (?, ?, 'Pendiente', ?, ?, ?, 'Producción', 'Espagiria', ?, 'Compra')
                    """, (candidato, fecha_now, user, urgencia, obs, categoria))
                    numero = candidato
                    break
                except sqlite3.IntegrityError as e:
                    ultimo_err = e
                    continue
            if numero is None:
                raise RuntimeError(
                    f"No se pudo asignar numero SOL tras 6 reintentos · {ultimo_err}"
                )

            items_count = 0
            total_g_mps = 0.0
            for mp in mps_lst:
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado, proveedor_sugerido)
                        VALUES (?, ?, ?, ?, 'g', ?, 0, ?)
                    """, (numero, mp['codigo_mp'], mp['nombre'],
                          mp['faltante_g'],
                          f"Falta para producción {dias}d · stock {mp['stock_actual_g']}g de {mp['necesario_total_g']}g necesarios",
                          prov_label))
                except sqlite3.OperationalError:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado)
                        VALUES (?, ?, ?, ?, 'g', ?, 0)
                    """, (numero, mp['codigo_mp'], mp['nombre'],
                          mp['faltante_g'],
                          f"Falta para producción {dias}d · stock {mp['stock_actual_g']}g de {mp['necesario_total_g']}g"))
                items_count += 1
                total_g_mps += float(mp.get('faltante_g') or 0)

            for me in mees_lst:
                # Items MEE · usar nombre_mp como descripcion + cantidad en unidades
                try:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado, proveedor_sugerido)
                        VALUES (?, ?, ?, ?, 'unidades', ?, 0, ?)
                    """, (numero, me['codigo'], me['descripcion'],
                          me['faltante_u'],
                          f"Falta para envasar producción {dias}d · stock {me['stock_actual_u']}u de {me['necesario_total_u']}u",
                          prov_label))
                except sqlite3.OperationalError:
                    c.execute("""
                        INSERT INTO solicitudes_compra_items
                          (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                           justificacion, valor_estimado)
                        VALUES (?, ?, ?, ?, 'unidades', ?, 0)
                    """, (numero, me['codigo'], me['descripcion'],
                          me['faltante_u'],
                          f"Falta para envasar {dias}d · stock {me['stock_actual_u']}u de {me['necesario_total_u']}u"))
                items_count += 1

            try:
                audit_log(
                    c, usuario=user, accion='SOLICITAR_FALTANTES_BULK',
                    tabla='solicitudes_compra', registro_id=numero,
                    despues={
                        'proveedor': prov_label or '(sin proveedor)',
                        'horizonte_dias': dias,
                        'urgencia': urgencia,
                        'mps_count': len(mps_lst),
                        'mees_count': len(mees_lst),
                        'items_count': items_count,
                    },
                    detalle=(f"Auto-bulk: {numero} · {prov_label or 'sin proveedor'} · "
                              f"{len(mps_lst)} MPs + {len(mees_lst)} MEEs"),
                )
            except Exception as _e:
                log.warning('audit_log SOLICITAR_FALTANTES_BULK fallo: %s', _e)

            creadas.append({
                'numero': numero,
                'proveedor': prov_label or '(sin proveedor)',
                'items_count': items_count,
                'mps_count': len(mps_lst),
                'mees_count': len(mees_lst),
                'total_g_mps': round(total_g_mps, 2),
            })

        conn.commit()
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        log.exception('solicitar_faltantes_bulk fallo: %s', e)
        return jsonify({'error': str(e)}), 500

    return jsonify({
        'ok': True,
        'mensaje': f'Creadas {len(creadas)} solicitudes para {len(creadas)} proveedores',
        'solicitudes_creadas': creadas,
        'total_proveedores': len(creadas),
        'total_solicitudes': len(creadas),
        'horizonte_dias': dias,
    }), 201


# ─── Planificación Estratégica ────────────────────────────────────────────────

@bp.route('/api/programacion/planificacion')
def planificacion_estrategica():
    """
    Proyección de MPs para 2, 6 o 12 meses basada en calendario de producción.
    Detecta déficits, oportunidades de compra en volumen e inteligencia de origen.
    """
    import datetime as _dt
    import re as _re

    # Aceptar ?dias=N (15/30/60/90/180/365) o ?meses=N (back-compat).
    # dias prevalece si esta presente.
    dias_param = request.args.get('dias')
    if dias_param:
        try:
            dias = max(1, min(int(dias_param), 365))
        except ValueError:
            dias = 60
        meses = max(1, dias // 31 + (1 if dias % 31 else 0))
    else:
        meses = min(int(request.args.get('meses', 2)), 12)
        dias = meses * 31
    days_ahead = dias  # cutoff exacto basado en dias del horizonte

    conn = get_db()

    # ── 1. Cargar datos base ──────────────────────────────────────────────────
    cal     = _fetch_calendar_events(days_ahead=days_ahead)
    events  = cal.get('events', [])
    formulas = _get_formulas(conn)
    mp_stock = _get_mp_stock(conn)

    # SKU → producto map
    _sku_to_prod = {}
    try:
        for row in conn.execute(
            "SELECT UPPER(sku), producto_nombre FROM sku_producto_map WHERE activo=1"
        ).fetchall():
            _sku_to_prod[row[0]] = row[1]
    except Exception:
        pass

    # Proveedor map: material_id → proveedor
    _prov_map = {}
    try:
        for row in conn.execute(
            "SELECT codigo_mp, proveedor FROM maestro_mps WHERE proveedor IS NOT NULL AND proveedor != ''"
        ).fetchall():
            _prov_map[row[0]] = row[1]
    except Exception:
        pass

    # ── 2. Parsear eventos del calendario ────────────────────────────────────
    today = _dt.date.today()
    cutoff = today + _dt.timedelta(days=days_ahead)

    _NOT_SKU = {
        'FAB','QC','CON','SIN','MICRO','ENVASADO','ACONDICIONAMIENTO',
        'FABRICACION','FABRICACIÓN','LANZAMIENTO','PRODUCCION','PRODUCCIÓN',
        'KG','MES','DIAS','DÍAS','ML','UDS','BATCH','FERNANDO',
        'MESA','LOTES','LOTE','MINI','MACRO','AND','THE','FOR'
    }

    def _skus(titulo):
        tokens = _re.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', titulo.upper())
        return [t for t in tokens if t not in _NOT_SKU]

    def _kg_ev(titulo):
        m = _re.findall(r'~?(\d+(?:[,.]\d+)*)\s*kg', titulo, _re.IGNORECASE)
        if not m: return None
        try: return float(m[-1].replace(',', '.'))
        except: return None

    # Producciones planificadas: lista de {fecha, producto, kg, mes_label}
    producciones = []
    seen_events = set()

    # ── 2a. Fuente primaria: Google Calendar / iCal ───────────────────────
    for ev in events:
        titulo   = ev.get('titulo', '')
        fecha_s  = ev.get('fecha', '')
        if not fecha_s: continue
        try:
            fecha = _dt.date.fromisoformat(fecha_s)
        except ValueError:
            continue
        if fecha < today or fecha > cutoff:
            continue

        key = (titulo.strip(), fecha_s)
        if key in seen_events:
            continue
        seen_events.add(key)

        # ── Phase filter: skip non-Fabricación events ──────────────────────
        # Envasado / Acondicionamiento / QC no consumen MPs crudas.
        # Contarlos duplica/triplica los requerimientos (infla 2-3×).
        # Sebastián 12-may-2026: usar constante global.
        if any(kw in titulo.lower() for kw in NON_FAB_KW_GLOBAL):
            continue

        kg = _kg_ev(titulo)
        for sku in _skus(titulo):
            prod = _sku_to_prod.get(sku)
            if prod and prod in formulas:
                mes_label = fecha.strftime('%Y-%m')
                producciones.append({
                    'fecha': fecha_s,
                    'mes':   mes_label,
                    'producto': prod,
                    'kg': kg or formulas[prod]['lote_size_kg'],
                    'sku': sku,
                    'titulo': titulo,
                    'fuente': 'calendario',
                })
                break

    # ── 2b. Fuente secundaria: produccion_programada (DB local, MANUALES) ─
    # Complementa cuando el calendario no tiene eventos o no hay SKUs en títulos.
    # IMPORTANTE: filtramos origen='calendar' (filas auto-sync) para evitar
    # duplicados — esas ya se procesan en 2a directamente del calendar source.
    # Solo manuales (origen='manual' o NULL).
    _upper_to_prod = {p.upper(): p for p in formulas.keys()}
    try:
        local_rows = conn.execute(
            """SELECT producto, fecha_programada, lotes FROM produccion_programada
               WHERE estado NOT IN ('completado','cancelado')
                 AND fecha_programada >= date('now', '-5 hours', '-7 days')
                 AND fecha_programada <= ?
                 AND COALESCE(origen,'manual') NOT IN ('calendar','eos_canonico')
               ORDER BY fecha_programada""",
            (cutoff.isoformat(),)
        ).fetchall()
        for row in local_rows:
            prod_raw = (row[0] or '').strip()
            fecha_s  = (row[1] or '').strip()
            lotes    = int(row[2] or 1)
            prod = prod_raw if prod_raw in formulas else _upper_to_prod.get(prod_raw.upper())
            if not prod or not fecha_s:
                continue
            key = (prod, fecha_s)
            if key in seen_events:
                continue
            seen_events.add(key)
            try:
                fecha = _dt.date.fromisoformat(fecha_s)
            except ValueError:
                continue
            lote_kg = formulas[prod]['lote_size_kg']
            mes_label = fecha.strftime('%Y-%m')
            producciones.append({
                'fecha': fecha_s,
                'mes': mes_label,
                'producto': prod,
                'kg': lote_kg * lotes,
                'sku': '',
                'titulo': f'{prod} — {lotes} lote(s)',
                'fuente': 'local',
            })
    except Exception:
        pass

    # ── 3. Calcular MPs necesarias por producción ────────────────────────────
    # mp_needed: {material_id: {nombre, total_g, meses: set(), prods: list}}
    mp_needed = {}

    # Helper de stock antes del loop para reusarlo en mps_status por-producción
    # Ahora con ROLLING STOCK: el stock se va decrementando a medida que las
    # producciones cronologicamente anteriores consumen MPs.
    # Bug fix 2026-04-28: antes evaluaba cada produccion contra el stock total,
    # asi que si dos producciones usaban la misma MP, ambas decian "puede"
    # aunque entre las dos no alcanzara el stock.
    def _lookup_stock_key(mid, nombre):
        """Devuelve la clave del dict mp_stock usada para este MP, o None
        si no se encuentra. Permite decrementar el stock simulado en la
        clave correcta despues."""
        if _is_unlimited_mp(nombre):
            return None  # ilimitado, no decrementar
        # FIX 1-jun-2026 audit Abastecimiento (P0) · incluir tier ALIAS (faltaba) ·
        # alinea el lookup de factibilidad con el canónico → MP nombre-variante
        # resuelve a su clave real y NO da déficit/factibilidad espuria.
        _ne = (nombre or '').upper()
        _nn = _norm_mp_name(nombre or '')
        _al = _MP_NAME_ALIAS.get(_nn) or _MP_NAME_ALIAS.get(_ne)
        for k in (mid, (mid or '').upper(), _ne, _nn, _al):
            if k and k in stock_simulado:
                return k
        return None

    def _lookup_stock(mid, nombre):
        """Devuelve stock simulado actual del MP. -1 si es ilimitado."""
        if _is_unlimited_mp(nombre):
            return -1
        k = _lookup_stock_key(mid, nombre)
        if k is None:
            return 0
        return stock_simulado[k]

    # Stock simulado: copia mutable del stock real que se decrementa a
    # medida que avanzan las producciones cronologicas.
    stock_simulado = dict(mp_stock)

    # ORDEN CRITICO: las producciones DEBEN evaluarse en orden cronologico
    # para que el rolling stock tenga sentido.
    producciones.sort(key=lambda p: (p.get('fecha', ''), p.get('producto', '')))

    for prod_ev in producciones:
        prod   = prod_ev['producto']
        kg_ev  = prod_ev['kg']
        mes    = prod_ev['mes']
        formula = formulas.get(prod, {})
        lote_kg = formula.get('lote_size_kg', 1)
        factor  = kg_ev / lote_kg if lote_kg > 0 else 1.0

        # Per-producción MP status — evalua contra stock_simulado actual.
        # Si esta produccion alcanza, decrementa stock_simulado para la
        # siguiente produccion cronologica.
        mps_status = []
        n_alcanza = 0
        n_falta   = 0
        decrementos_si_alcanza = []  # (key, g_need) para aplicar al final

        for item in formula.get('items', []):
            mid    = item['material_id']
            nombre = item['material_nombre']
            g_lote = float(item.get('cantidad_g_por_lote', 0) or 0)
            pct    = float(item.get('porcentaje', 0) or 0)
            # ABASTECIMIENTO-FIX · 22-may-2026 · fallback porcentaje (#5 audit 22-may)
            # · Si fórmula tiene porcentaje pero cant_lote=0, calcular con %
            #   asume formula.lote_size_kg como base · sino usa factor directo
            g_need = round(g_lote * factor, 1)
            if g_need <= 0 and pct > 0:
                # Fallback: pct% del batch (kg de la producción × 1000 × %/100)
                kg_batch = formula.get('lote_size_kg') or formula.get('kg_total') or 1
                g_need = round(float(kg_batch) * 1000 * pct / 100 * factor, 1)
            if g_need <= 0:
                continue

            if mid not in mp_needed:
                mp_needed[mid] = {
                    'material_id': mid,
                    'nombre': nombre,
                    'total_g': 0,
                    'por_mes': {},   # mes_label → g
                    'productos': [],
                }
            mp_needed[mid]['total_g'] += g_need
            mp_needed[mid]['por_mes'][mes] = mp_needed[mid]['por_mes'].get(mes, 0) + g_need
            if prod not in mp_needed[mid]['productos']:
                mp_needed[mid]['productos'].append(prod)

            # Status individual para esta producción usando stock SIMULADO
            stock_g_raw = _lookup_stock(mid, nombre)
            stock_key   = _lookup_stock_key(mid, nombre)

            if stock_g_raw == -1:
                # Ilimitado (agua, etc.) — siempre alcanza, no decrementa
                mps_status.append({
                    'material_id': mid, 'nombre': nombre,
                    'necesario_g': g_need, 'stock_g': -1,
                    'alcanza': True, 'deficit_g': 0,
                    'ilimitado': True,
                })
                n_alcanza += 1
            else:
                alcanza = stock_g_raw >= g_need
                mps_status.append({
                    'material_id': mid, 'nombre': nombre,
                    'necesario_g': g_need, 'stock_g': round(stock_g_raw, 1),
                    'alcanza': alcanza,
                    'deficit_g': round(max(0, g_need - stock_g_raw), 1),
                    'ilimitado': False,
                })
                if alcanza:
                    n_alcanza += 1
                    if stock_key is not None:
                        decrementos_si_alcanza.append((stock_key, g_need))
                else:
                    n_falta += 1

        # Solo decrementar stock_simulado si TODAS las MPs alcanzan
        # (de lo contrario la produccion no se va a hacer y no se consume nada).
        puede_producir = (n_falta == 0)
        if puede_producir:
            for k, g in decrementos_si_alcanza:
                stock_simulado[k] = max(0, stock_simulado.get(k, 0) - g)

        prod_ev['mps_status']    = mps_status
        prod_ev['n_mps_alcanza'] = n_alcanza
        prod_ev['n_mps_falta']   = n_falta
        prod_ev['puede_producir'] = puede_producir

    # ── 4. Cruzar con stock actual → déficit ────────────────────────────────

    resultado = []
    for mid, data in mp_needed.items():
        stock_g_raw = _lookup_stock(mid, data['nombre'])
        if stock_g_raw == -1:
            # Producido en sitio (agua desionizada, etc.) — sin déficit, sin compra
            continue
        stock_g   = stock_g_raw
        deficit   = max(0, data['total_g'] - stock_g)
        cobertura = round(min(stock_g / data['total_g'] * 100, 100), 1) if data['total_g'] > 0 else 100
        meses_uso = sorted(data['por_mes'].keys())
        n_meses   = len(meses_uso)
        proveedor = _prov_map.get(mid, '')

        # ── Inteligencia de origen ──────────────────────────────────────────
        # Palabras clave que sugieren origen China / importación directa
        _CHINA_KEYWORDS = {'lyphar','tianki','bloomage','sinomax','croda','basf','evonik',
                           'givaudan','dsm','lubrizol','ashland','clariant','solvay',
                           'chinese','china','guangzhou','shanghai','beijing'}
        _COL_KEYWORDS   = {'inchemical','en qu','quiminet','prodycon','corquiven','quimicos',
                           'colombia','agenquimicos','ytbio','bolite','laboratorios','bogota',
                           'medellin','cali','distribuidora'}
        prov_lower = proveedor.lower()
        is_china   = any(k in prov_lower for k in _CHINA_KEYWORDS) or mid.startswith('MP001') or mid.startswith('MP002')
        is_col     = any(k in prov_lower for k in _COL_KEYWORDS)

        # Oportunidad de bulk: si se usa en 3+ meses consecutivos
        bulk_opp = False
        bulk_msg = ''
        if n_meses >= 2:
            bulk_opp = True
            _qty = f'{int(round(data["total_g"])):,} g'.replace(',', '.')
            if is_china:
                bulk_msg = f'Importar {_qty} en un solo pedido desde proveedor internacional — ahorro estimado 15-25% en flete'
            else:
                bulk_msg = f'Pedir {_qty} para {n_meses} meses a proveedor local — negociar descuento por volumen'

        resultado.append({
            'material_id':  mid,
            'nombre':       data['nombre'],
            'proveedor':    proveedor,
            'total_g':      round(data['total_g'], 1),
            'stock_g':      round(stock_g, 1),
            'deficit_g':    round(deficit, 1),
            'cobertura_pct': cobertura,
            'meses_uso':    meses_uso,
            'n_meses':      n_meses,
            'productos':    data['productos'],
            'bulk_opp':     bulk_opp,
            'bulk_msg':     bulk_msg,
            'origen':       'china' if is_china else ('colombia' if is_col else 'desconocido'),
        })

    # Ordenar: déficit primero, luego por total requerido
    resultado.sort(key=lambda x: (-x['deficit_g'], -x['total_g']))

    # ── 5. Resumen ejecutivo ─────────────────────────────────────────────────
    total_prods   = len(producciones)
    total_mps     = len(resultado)
    mps_deficit   = [r for r in resultado if r['deficit_g'] > 0]
    mps_ok        = [r for r in resultado if r['deficit_g'] == 0]
    bulk_opps     = [r for r in resultado if r['bulk_opp'] and r['deficit_g'] > 0]
    meses_unicos  = sorted(set(p['mes'] for p in producciones))

    horizonte_label = (f'{dias} días' if dias <= 90 else
                       f'{round(dias/30)} meses' if dias < 365 else '1 año')

    return jsonify({
        'meses':           meses,
        'dias':            dias,
        'horizonte_label': horizonte_label,
        'producciones':    producciones,
        'meses_unicos':    meses_unicos,
        'total_prods':     total_prods,
        'total_mps':       total_mps,
        'mps_deficit':     mps_deficit,
        'mps_ok':          mps_ok,           # staff general view
        'mps_ok_count':    len(mps_ok),
        'bulk_opps':       bulk_opps,
        'cal_error':       cal.get('error'),
        'generado_en':     _dt.datetime.now().isoformat(),
    })


@bp.route('/api/programacion/planificacion/solicitar-bulk', methods=['POST'])
def planificacion_solicitar_bulk():
    """Crea solicitudes_compra agrupadas por proveedor para los MPs en
    deficit del horizonte indicado.

    Una solicitud por proveedor con todos sus MPs faltantes. Texto
    auto-generado deja claro que fue creada por planificacion bulk.

    Body:
      dias: int (15/30/60/90/180/365)
      urgencia: 'Alta'|'Normal'|'Baja' (default 'Normal')

    Reusa la logica del GET planificacion para no duplicar matemathics.

    Solo autenticado (cualquier user puede crear solicitud — es flujo
    operativo). Audit log captura accion.
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '') or 'sistema'

    d = request.json or {}
    try:
        dias = max(1, min(int(d.get('dias', 60)), 365))
    except ValueError:
        dias = 60
    urgencia = (d.get('urgencia') or 'Normal').strip()
    if urgencia not in ('Alta', 'Normal', 'Baja'):
        urgencia = 'Normal'

    # Llamar el endpoint de planificacion via test_request_context — pero
    # mejor reusar la logica directa para no enredar.
    from flask import current_app
    with current_app.test_request_context(f'/api/programacion/planificacion?dias={dias}'):
        plan_resp = planificacion_estrategica()
    plan_data = plan_resp.get_json()
    if not plan_data:
        return jsonify({'error': 'No pude obtener planificacion'}), 500

    deficits = plan_data.get('mps_deficit', []) or []
    # ABASTECIMIENTO-FIX · 22-may-2026 · dedup con cola pendiente (#9 audit 22-may)
    # · Antes: solicitar-bulk insertaba déficit completo aun si ya había SOLs Pendientes
    # · Ahora: para cada MP resta _pendiente_en_compras_g · si neto <= 0 skip MP
    try:
        from blueprints.compras import _pendiente_en_compras_g
        conn = get_db()
        cur = conn.cursor()
        deficits_neto = []
        for mp in deficits:
            cod_mp = mp.get('material_id') or mp.get('codigo_mp', '')
            if not cod_mp:
                deficits_neto.append(mp)
                continue
            try:
                en_cola = _pendiente_en_compras_g(cur, cod_mp)
            except Exception:
                en_cola = 0
            deficit_g = float(mp.get('deficit_g') or 0)
            if en_cola >= deficit_g:
                continue  # ya cubierto · skip
            mp_copy = dict(mp)
            mp_copy['deficit_g'] = max(0, deficit_g - en_cola)
            mp_copy['en_cola_g'] = en_cola
            deficits_neto.append(mp_copy)
        deficits = deficits_neto
    except Exception:
        pass  # graceful · sigue con déficits crudos
    if not deficits:
        return jsonify({
            'ok': True,
            'mensaje': 'Sin déficits netos — todo ya está en cola de compras.',
            'solicitudes_creadas': [],
        })

    # Agrupar por proveedor canónico (el de maestro_mps)
    por_proveedor = {}
    for mp in deficits:
        prov = (mp.get('proveedor') or '(SIN PROVEEDOR)').strip() or '(SIN PROVEEDOR)'
        por_proveedor.setdefault(prov, []).append(mp)

    conn = get_db()
    c = conn.cursor()
    from datetime import datetime as _dt

    solicitudes_creadas = []
    errores = []
    for prov, mps in por_proveedor.items():
        try:
            # Generar numero
            year = _dt.now().strftime('%Y')
            row = c.execute(
                "SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) "
                "FROM solicitudes_compra WHERE numero LIKE ?",
                (f'SOL-{year}-%',)
            ).fetchone()
            num = (row[0] or 0) + 1
            numero = f'SOL-{year}-{num:04d}'

            obs = (f'Auto-generada Planificación Estratégica {dias} días — '
                   f'Proveedor: {prov} · {len(mps)} MPs · ' +
                   ', '.join([
                       f'{m["nombre"]} ({int(round(m["deficit_g"])):,} g)'.replace(',', '.')
                       for m in mps[:5]
                   ]) +
                   (f' +{len(mps)-5} más' if len(mps) > 5 else ''))[:1500]

            c.execute(
                """INSERT INTO solicitudes_compra
                   (numero, fecha, estado, solicitante, urgencia, observaciones,
                    area, empresa, categoria, tipo)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (numero, _dt.now().isoformat(), 'Pendiente', user, urgencia,
                 obs, 'Produccion', 'Espagiria', 'Materia Prima', 'Compra')
            )
            for mp in mps:
                c.execute(
                    """INSERT INTO solicitudes_compra_items
                       (numero, codigo_mp, nombre_mp, cantidad_g, unidad,
                        justificacion, valor_estimado)
                       VALUES (?,?,?,?,?,?,?)""",
                    (numero, mp.get('material_id', ''), mp.get('nombre', ''),
                     float(mp.get('deficit_g', 0)), 'g',
                     f'Déficit horizonte {dias}d (Planificación Estratégica)',
                     0)
                )
            conn.commit()
            solicitudes_creadas.append({
                'numero': numero,
                'proveedor': prov,
                'count_mps': len(mps),
                'total_g': round(sum(m.get('deficit_g', 0) for m in mps), 1),
            })
        except Exception as _e:
            try: conn.rollback()
            except Exception as _r:
                logging.getLogger('programacion').debug('rollback no aplicable: %s', _r)
            errores.append({'proveedor': prov, 'error': str(_e)})
            continue

    # Audit
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (user, 'PLANIFICACION_BULK_REQUEST', 'solicitudes_compra',
                   '_BULK_',
                   _json.dumps({
                       'dias': dias, 'urgencia': urgencia,
                       'count_solicitudes': len(solicitudes_creadas),
                       'count_errores': len(errores),
                       'total_g_pedido': round(sum(
                           s['total_g'] for s in solicitudes_creadas
                       ), 1),
                   }, ensure_ascii=False),
                   request.remote_addr))
        conn.commit()
    except sqlite3.OperationalError:
        pass

    return jsonify({
        'ok': True,
        'horizonte_dias': dias,
        'count_solicitudes': len(solicitudes_creadas),
        'count_errores': len(errores),
        'solicitudes_creadas': solicitudes_creadas,
        'errores': errores,
        'mensaje': (
            f'{len(solicitudes_creadas)} solicitudes creadas '
            f'(una por proveedor) para horizonte {dias} días.'
        ),
    })


@bp.route('/api/programacion/planificacion/checklist-verificacion')
def planificacion_checklist_verificacion():
    """Genera XLSX con MPs reportadas en STOCK CERO por horizonte para que
    la asistente verifique fisicamente en bodega si estan o no.

    Query params:
      horizontes: csv de dias (default '15,30'). Cada uno es una hoja.

    Cada hoja incluye:
      - Material, Codigo, Proveedor, Necesario (g/kg), Para Producto(s)
      - Casillas vacias: 'En bodega? (S/N)', 'Cantidad real (g)', 'Notas'
      - Solo MPs con stock_g <= 0 (las que parcialmente tienen stock no
        van — ya sabemos que existen, solo falta cantidad)

    Auth requerida.
    """
    if not _auth():
        return jsonify({'error': 'No autenticado'}), 401

    horizontes_param = request.args.get('horizontes', '15,30')
    horizontes = []
    for h in horizontes_param.split(','):
        h = h.strip()
        if not h: continue
        try:
            v = max(1, min(int(h), 365))
            if v not in horizontes:
                horizontes.append(v)
        except ValueError:
            continue
    if not horizontes:
        horizontes = [15, 30]

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl no disponible'}), 500

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    from flask import current_app, send_file
    import io

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill   = PatternFill('solid', fgColor='1A4A7A')
    title_fill    = PatternFill('solid', fgColor='1A4A7A')
    alt_fill      = PatternFill('solid', fgColor='F8F9FA')
    warn_fill     = PatternFill('solid', fgColor='FFF3CD')

    total_en_cero = 0
    for dias in horizontes:
        with current_app.test_request_context(f'/api/programacion/planificacion?dias={dias}'):
            plan_resp = planificacion_estrategica()
        plan_data = plan_resp.get_json() or {}
        deficits = plan_data.get('mps_deficit', []) or []
        en_cero = [mp for mp in deficits if (mp.get('stock_g') or 0) <= 0]
        # Ordenar por necesario desc
        en_cero.sort(key=lambda m: -(m.get('total_g') or 0))
        total_en_cero += len(en_cero)

        # Etiqueta de hoja: '15 dias' / '1 mes' / '60 dias' etc
        if dias == 15:    sheet_name = '15 dias'
        elif dias == 30:  sheet_name = '1 mes'
        elif dias == 60:  sheet_name = '2 meses'
        elif dias == 90:  sheet_name = '3 meses'
        elif dias == 180: sheet_name = '6 meses'
        elif dias == 365: sheet_name = '1 ano'
        else:             sheet_name = f'{dias} dias'

        ws = wb.create_sheet(sheet_name)

        # Titulo
        ws.cell(row=1, column=1, value=f'Lista de verificacion de bodega — Horizonte {sheet_name}')
        ws.cell(row=1, column=1).font = Font(size=14, bold=True, color='FFFFFF')
        ws.cell(row=1, column=1).fill = title_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.row_dimensions[1].height = 24

        # Subtitulo
        gen = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.cell(row=2, column=1,
                value=f'Generado: {gen} · {len(en_cero)} MP(s) reportada(s) en STOCK CERO por sistema · Verificar fisicamente en bodega y marcar.')
        ws.cell(row=2, column=1).font = Font(size=10, italic=True, color='666666')
        ws.cell(row=2, column=1).fill = warn_fill
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        ws.row_dimensions[2].height = 30

        # Nota: hint sobre duplicados
        ws.cell(row=3, column=1,
                value='OJO: Si un material no esta bajo el codigo indicado, busca por NOMBRE — puede estar bajo codigo distinto. Anota lo que encuentres en "Notas".')
        ws.cell(row=3, column=1).font = Font(size=9, color='856404')
        ws.cell(row=3, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
        ws.row_dimensions[3].height = 20

        # Headers — todo en gramos (acordado con Alejandro)
        headers = [
            'Material', 'Codigo', 'Proveedor',
            'Necesario (g)',
            'Para producto(s)',
            'En bodega? (S/N)', 'Cantidad real (g)',
        ]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[5].height = 32

        # Filas
        if en_cero:
            for idx, mp in enumerate(en_cero, 6):
                nec_g = float(mp.get('total_g') or 0)
                values = [
                    mp.get('nombre', ''),
                    mp.get('material_id', ''),
                    mp.get('proveedor', '') or '(Sin asignar)',
                    int(round(nec_g)),
                    ', '.join(mp.get('productos', []) or []),
                    '',  # asistente marca S/N
                    '',  # cantidad real
                ]
                for j, v in enumerate(values, 1):
                    cell = ws.cell(row=idx, column=j, value=v)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal='right' if j == 4 else ('center' if j == 6 else 'left'),
                        vertical='center', wrap_text=(j == 5),
                    )
                    if j == 4:  # cantidad necesaria con separador de miles
                        cell.number_format = '#,##0'
                    if (idx - 6) % 2 == 1:
                        cell.fill = alt_fill
                    if j == 2:  # codigo monoespaciado
                        cell.font = Font(name='Consolas', size=9)
                # Resaltar columnas de verificacion
                ws.cell(row=idx, column=6).fill = warn_fill
                ws.cell(row=idx, column=7).fill = warn_fill
        else:
            ws.cell(row=6, column=1, value='Sin MPs en stock cero para este horizonte')
            ws.cell(row=6, column=1).font = Font(italic=True, color='888888')
            ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=7)

        # Anchos columnas
        widths = [30, 16, 18, 14, 42, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header
        ws.freeze_panes = 'A6'

    # Audit
    try:
        conn = get_db()
        c = conn.cursor()
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (session.get('compras_user', '') or 'sistema',
                   'PLANIFICACION_CHECKLIST_DOWNLOAD',
                   'planificacion', '_CHECKLIST_',
                   _json.dumps({'horizontes': horizontes, 'total_en_cero': total_en_cero},
                               ensure_ascii=False),
                   request.remote_addr))
        conn.commit()
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'verificar_bodega_{"-".join(str(h) for h in horizontes)}d_{date.today().isoformat()}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


# ═════════════════════════════════════════════════════════════════════════
#  CHECKLIST PRE-PRODUCCION
#  Sistema que para cada produccion programada genera lista de items a
#  verificar con anticipacion: MPs, envases, tapas, etiquetas, serigrafia,
#  tampografia. Conecta con compras: si falta algo, genera SOL automatica.
# ═════════════════════════════════════════════════════════════════════════

ITEMS_CHECKLIST_DEFAULT = [
    ('envase_primario',  'Envase primario (frasco/contenedor)', 30, 1),
    ('tapa',             'Tapa o sistema dosificador',          30, 2),
    ('etiqueta_frontal', 'Etiqueta frontal',                    25, 3),
    ('etiqueta_posterior','Etiqueta posterior con info legal',  25, 4),
    ('caja_exterior',    'Caja exterior individual',            20, 5),
    ('serigrafia',       'Serigrafia en envase si aplica',      30, 6),
    ('tampografia',      'Tampografia en tapa si aplica',       30, 7),
]

# Tipos legacy que ya NO se generan ni muestran. La decoracion del envase
# (etiqueta_adhesiva / serigrafia / tampografia) cubre estas necesidades
# y crea las OCs/tareas operativas correspondientes desde Compras.
# Pedido Sebastian 2026-04-29: "quita lo de etiquetas frontales porque ya
# esta en el que sale". Si necesitas reactivarlas, vacia esta tupla.
TIPOS_LEGACY_OCULTOS = ('etiqueta_frontal', 'etiqueta_posterior', 'etiqueta_lateral')


def _calcular_disponibilidad_mp(c, codigo_mp, fecha_horizonte=None):
    """Calcula disponibilidad real de un MP considerando TODAS las
    producciones programadas hasta fecha_horizonte.

    Returns dict con:
      - stock_actual: gramos en bodega ahora (sumando movimientos)
      - en_transito: gramos en OCs activas no recibidas todavia
                     (Aprobada/Autorizada/Pagada categoria MP)
      - demanda_total_horizonte: gramos requeridos para TODAS las
                                 producciones programadas hasta fecha
      - solicitado_pendiente: gramos en SOLs activas (no convertidas a OC)
      - disponibilidad_neta: stock + en_transito - demanda
                             (si negativo, hay que solicitar)

    Esto resuelve la queja de Sebastian: el checklist NO debe ver una
    produccion en aislado, sino TODAS las del horizonte. Si la suma de
    demanda excede stock+transito, hay que pedir HOY aunque sea para
    una produccion de dentro de un mes.
    """
    if not codigo_mp:
        return {'stock_actual': 0, 'en_transito': 0,
                'demanda_total_horizonte': 0,
                'solicitado_pendiente': 0,
                'disponibilidad_neta': 0}
    out = {}
    # 1) Stock actual
    try:
        row = c.execute("""
            SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','Ajuste +') THEN cantidad
                                     WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad
                                     ELSE 0 END), 0)
            FROM movimientos WHERE material_id=?
        """, (codigo_mp,)).fetchone()
        out['stock_actual'] = float(row[0] or 0)
    except Exception:
        out['stock_actual'] = 0.0

    # 2) En transito: OCs en estados activos no Recibida/Pagada
    # Sumamos cantidad_g de items de OCs en Borrador/Pendiente/Revisada/
    # Aprobada/Autorizada/Parcial. Pagada significa que se pago pero la
    # mercancia puede no haber llegado todavia, asi que tambien suma.
    try:
        row = c.execute("""
            SELECT COALESCE(SUM(i.cantidad_g - COALESCE(i.cantidad_recibida_g,0)), 0)
            FROM ordenes_compra_items i
            JOIN ordenes_compra oc ON oc.numero_oc = i.numero_oc
            WHERE i.codigo_mp = ?
              AND oc.estado IN ('Borrador','Pendiente','Revisada','Aprobada',
                                'Autorizada','Parcial','Pagada')
        """, (codigo_mp,)).fetchone()
        out['en_transito'] = float(row[0] or 0)
    except Exception:
        out['en_transito'] = 0.0

    # 3) En SOL pendientes (aun no convertidas a OC)
    try:
        row = c.execute("""
            SELECT COALESCE(SUM(si.cantidad_g), 0)
            FROM solicitudes_compra_items si
            JOIN solicitudes_compra s ON s.numero = si.numero
            WHERE si.codigo_mp = ?
              AND s.estado IN ('Pendiente','En revision','Aprobada')
              AND COALESCE(s.numero_oc, '') = ''
        """, (codigo_mp,)).fetchone()
        out['solicitado_pendiente'] = float(row[0] or 0)
    except Exception:
        out['solicitado_pendiente'] = 0.0

    # 4) Demanda total del horizonte: sumar de TODAS las producciones
    # programadas hasta fecha_horizonte cuyo formula_items incluya este MP.
    # produccion_programada tiene cols: producto, fecha_programada, lotes
    # cantidad_g real = fi.cantidad_g_por_lote * pp.lotes
    try:
        params = [codigo_mp]
        sql = """
            SELECT COALESCE(SUM(
                COALESCE(fi.cantidad_g_por_lote, 0) * COALESCE(pp.lotes, 1)
            ), 0) as demanda_g
            FROM produccion_programada pp
            JOIN formula_items fi ON fi.producto_nombre = pp.producto
            WHERE fi.material_id = ?
              AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado')
              AND pp.fecha_programada >= date('now', '-5 hours', '-1 day')
        """
        if fecha_horizonte:
            sql += " AND pp.fecha_programada <= ?"
            params.append(fecha_horizonte)
        row = c.execute(sql, params).fetchone()
        out['demanda_total_horizonte'] = float(row[0] or 0)
    except Exception:
        out['demanda_total_horizonte'] = 0.0

    # 5) Disponibilidad neta = stock + transito + solicitado - demanda
    # solicitado_pendiente todavia no es seguro pero suma como cobertura tentativa.
    out['disponibilidad_neta'] = (
        out['stock_actual']
        + out['en_transito']
        + out['solicitado_pendiente']
        - out['demanda_total_horizonte']
    )
    return out


def _generar_checklist_produccion(c, produccion_id, producto_nombre, fecha_planeada,
                                   cantidad_kg, generar_mps=True, usuario='sistema'):
    """Genera items de checklist para una produccion.

    Si generar_mps=True, lee la formula del producto y crea 1 item por cada
    MP con la cantidad requerida + analisis de disponibilidad CONSIDERANDO
    TODAS las producciones del horizonte (decision Sebastian 2026-04-28).

    Tambien crea items default de envases/etiquetas/etc segun plantilla.

    Idempotente: si ya hay items para esta produccion_id, no duplica.
    """
    # Verificar si ya hay items
    existing = c.execute(
        "SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=?",
        (produccion_id,)
    ).fetchone()[0]
    if existing > 0:
        return 0  # ya tiene checklist

    items_creados = 0

    # 1) MPs desde la formula con calculo de disponibilidad agregada
    if generar_mps:
        try:
            mps = c.execute("""
                SELECT material_id, material_nombre, porcentaje, cantidad_g_por_lote
                FROM formula_items
                WHERE producto_nombre = ?
            """, (producto_nombre,)).fetchall()
            for mp in mps:
                codigo_mp = mp[0] or ''
                nombre_mp = mp[1] or ''
                porcentaje = float(mp[2] or 0)
                # Cantidad requerida SOLO para esta produccion individual
                cant_req_g = (porcentaje / 100.0) * (cantidad_kg or 0) * 1000.0

                # Disponibilidad agregada considerando TODO el horizonte
                # hasta la fecha de esta produccion (inclusive)
                disp = _calcular_disponibilidad_mp(c, codigo_mp,
                                                    fecha_horizonte=fecha_planeada)

                # Determinar estado segun disponibilidad NETA del horizonte
                if disp['disponibilidad_neta'] >= 0 and disp['en_transito'] == 0 and disp['solicitado_pendiente'] == 0:
                    # Stock cubre la demanda total
                    estado = 'verificado_ok'
                elif disp['en_transito'] > 0 and disp['disponibilidad_neta'] >= 0:
                    # Hay material en transito que cubre o complementa
                    estado = 'en_transito'
                elif disp['solicitado_pendiente'] > 0 and disp['disponibilidad_neta'] >= 0:
                    # Hay SOL pendiente que cubre
                    estado = 'solicitado'
                else:
                    # Falta material: debe solicitarse HOY
                    estado = 'pendiente'

                # Deficit = lo que aun falta cubrir
                deficit = max(0, -disp['disponibilidad_neta'])

                obs_extra = (
                    f"Stock: {disp['stock_actual']:.0f}g "
                    f"+ Transito: {disp['en_transito']:.0f}g "
                    f"+ Solicitado: {disp['solicitado_pendiente']:.0f}g "
                    f"- Demanda total horizonte: {disp['demanda_total_horizonte']:.0f}g "
                    f"= Neta: {disp['disponibilidad_neta']:.0f}g"
                )

                c.execute("""INSERT INTO produccion_checklist
                    (produccion_id, producto_nombre, fecha_planeada, cantidad_kg,
                     item_tipo, descripcion, cantidad_requerida, unidad, codigo_mp,
                     stock_actual, deficit, estado, dias_anticipacion,
                     observaciones, actualizado_por)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (produccion_id, producto_nombre, fecha_planeada, cantidad_kg,
                     'mp', nombre_mp, cant_req_g, 'g', codigo_mp,
                     disp['stock_actual'], deficit, estado, 30,
                     obs_extra, usuario))
                items_creados += 1
        except Exception:
            pass

    # 2) Envases / etiquetas / decoracion (plantilla del producto o default)
    # Sebastian (29-abr-2026): "ya la app sabe la presentacion del producto,
    # deberia calcular en automatico la cantidad de envases ... pero con la
    # opcion de corregir de ser necesario". Calculamos cantidad_unidades
    # desde formula_headers.volumen_unitario_ml o peso_g; el editor inline
    # ya existente permite ajustar manualmente.
    try:
        # Presentacion del producto (en g o ml — para cosmeticos densidad ~1)
        presentacion_g_ml = 0.0
        try:
            pres = c.execute("""
                SELECT COALESCE(volumen_unitario_ml, 0), COALESCE(peso_g, 0)
                FROM formula_headers
                WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
            """, (producto_nombre,)).fetchone()
            if pres:
                vol_ml = float(pres[0] or 0)
                peso_g = float(pres[1] or 0)
                # Preferir volumen_unitario_ml; fallback a peso_g de Shopify
                presentacion_g_ml = vol_ml or peso_g
        except Exception:
            pass

        # Calcular unidades objetivo con 5% de merma (estandar industrial)
        import math as _math
        if presentacion_g_ml > 0 and (cantidad_kg or 0) > 0:
            unidades_objetivo = int(_math.ceil(
                (cantidad_kg * 1000.0) / presentacion_g_ml * 1.05
            ))
        else:
            unidades_objetivo = 0  # sin info; frontend mostrara "—"

        # Tipos que llevan cantidad_unidades (los que son piezas fisicas);
        # serigrafia/tampografia son servicios, no aplican unidades.
        TIPOS_CON_UNIDADES = {
            'envase_primario', 'envase_secundario', 'tapa', 'caja_exterior',
            'etiqueta_frontal', 'etiqueta_posterior', 'etiqueta_lateral',
            'instructivo', 'otro',
        }

        # Buscar plantilla especifica del producto
        plantilla = c.execute("""
            SELECT item_tipo, descripcion, dias_anticipacion, orden, proveedor_default, obligatorio
            FROM checklist_plantillas
            WHERE producto_nombre=? OR producto_nombre=''
            ORDER BY CASE WHEN producto_nombre=? THEN 0 ELSE 1 END, orden
        """, (producto_nombre, producto_nombre)).fetchall()

        # Deduplicar: prioritar la del producto sobre la generica
        seen = set()
        items_a_crear = []
        for row in plantilla:
            tipo = row[0]
            if tipo in seen:
                continue
            seen.add(tipo)
            items_a_crear.append(row)

        for row in items_a_crear:
            tipo, desc, dias, orden, prov, obligatorio = row
            if not obligatorio:
                continue  # solo crear los obligatorios automaticamente
            if tipo in TIPOS_LEGACY_OCULTOS:
                continue  # cubierto por la decoracion del envase
            cant_ud = unidades_objetivo if tipo in TIPOS_CON_UNIDADES else 0
            unidad_label = 'ud' if cant_ud > 0 else ''
            obs_calc = (
                f"Auto: {cantidad_kg:.1f}kg / {presentacion_g_ml:.0f}{'ml' if (pres and (pres[0] or 0)>0) else 'g'} "
                f"= {cant_ud} ud (+5% merma)"
            ) if cant_ud > 0 else ''
            c.execute("""INSERT INTO produccion_checklist
                (produccion_id, producto_nombre, fecha_planeada, cantidad_kg,
                 item_tipo, descripcion, cantidad_unidades, unidad,
                 estado, proveedor, dias_anticipacion, observaciones,
                 actualizado_por)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (produccion_id, producto_nombre, fecha_planeada, cantidad_kg,
                 tipo, desc, cant_ud, unidad_label,
                 'pendiente', prov or '', dias or 30, obs_calc, usuario))
            items_creados += 1
    except Exception:
        pass

    return items_creados


@bp.route('/api/programacion/checklist/generar/<int:produccion_id>', methods=['POST'])
def checklist_generar(produccion_id):
    """Genera o regenera checklist para una produccion programada.

    Body opcional: {forzar: true} para regenerar borrando items existentes.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    conn = get_db(); c = conn.cursor()

    # Obtener datos de la produccion programada
    # produccion_programada tiene: producto, fecha_programada, lotes
    # kg total = lotes * formula_headers.lote_size_kg
    prod = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
               COALESCE(fh.lote_size_kg, 0) as lote_kg,
               COALESCE(pp.cantidad_kg, 0)  as cantidad_kg_explicita
        FROM produccion_programada pp
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
        WHERE pp.id=?
    """, (produccion_id,)).fetchone()
    if not prod:
        return jsonify({'error': 'Produccion no encontrada'}), 404
    lotes = int(prod[3] or 1)
    lote_kg = float(prod[4] or 0)
    cant_explicita = float(prod[5] or 0)
    # Prioridad: cantidad_kg explicita (del calendario) > lotes * lote_kg
    cant = cant_explicita if cant_explicita > 0 else (lotes * lote_kg)

    if d.get('forzar'):
        c.execute("DELETE FROM produccion_checklist WHERE produccion_id=?", (produccion_id,))

    items = _generar_checklist_produccion(
        c, produccion_id, prod[1], prod[2], cant,
        generar_mps=True,
        usuario=session.get('compras_user', 'sistema'))
    conn.commit()
    return jsonify({'ok': True, 'items_creados': items})


@bp.route('/api/programacion/checklist/<int:produccion_id>', methods=['GET'])
def checklist_get(produccion_id):
    """Devuelve el checklist completo de una produccion con totales por estado.

    Sebastian (28-abr-2026): el modal NO debe mostrar items tipo 'mp'
    (materias primas) — esos ya viven en Planificacion Estrategica con
    detalle agregado de stock+transito+solicitado-demanda. El checklist
    Pre-Produccion se enfoca en MEE: envases, tapas, etiquetas, serigrafia,
    tampografia. Por default, ?include_mps=0 (excluye MPs); pasar
    ?include_mps=1 para listarlas (uso debug/auditor).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    include_mps = request.args.get('include_mps', '0') == '1'
    conn = get_db(); c = conn.cursor()
    where_extra = "" if include_mps else " AND item_tipo != 'mp'"
    # Excluir tipos legacy que ya se cubren con la decoracion del envase
    if TIPOS_LEGACY_OCULTOS:
        placeholders = ','.join(['?'] * len(TIPOS_LEGACY_OCULTOS))
        where_extra += f" AND item_tipo NOT IN ({placeholders})"
        legacy_params = list(TIPOS_LEGACY_OCULTOS)
    else:
        legacy_params = []
    rows = c.execute(f"""
        SELECT id, item_tipo, descripcion, cantidad_requerida, unidad, codigo_mp,
               stock_actual, deficit, estado, proveedor, solicitud_numero,
               oc_numero, fecha_solicitud, fecha_eta, fecha_recibido,
               responsable, observaciones, dias_anticipacion, actualizado_at,
               producto_nombre, fecha_planeada
        FROM produccion_checklist
        WHERE produccion_id=?{where_extra}
        ORDER BY
          CASE item_tipo WHEN 'mp' THEN 1
                         WHEN 'envase_primario' THEN 2
                         WHEN 'tapa' THEN 3
                         WHEN 'etiqueta_frontal' THEN 4
                         WHEN 'etiqueta_posterior' THEN 5
                         WHEN 'caja_exterior' THEN 6
                         WHEN 'serigrafia' THEN 7
                         WHEN 'tampografia' THEN 8
                         ELSE 9 END,
          descripcion ASC
    """, [produccion_id] + legacy_params).fetchall()
    cols = [x[0] for x in c.description]
    items = [dict(zip(cols, r)) for r in rows]

    # Totales por estado (solo de los items mostrados — sin MPs por default)
    totales = {'pendiente': 0, 'verificado_ok': 0, 'solicitado': 0,
               'en_transito': 0, 'recibido': 0, 'listo': 0, 'no_aplica': 0}
    for it in items:
        totales[it['estado']] = totales.get(it['estado'], 0) + 1

    total_items = len(items)
    completados = totales['verificado_ok'] + totales['recibido'] + totales['listo']
    porcentaje = round((completados / total_items * 100), 1) if total_items > 0 else 0

    # Metadata del producto desde formula_headers + auto-sync Shopify si pendiente
    producto_nombre = ''
    producto_meta = {}
    try:
        if items and items[0].get('producto_nombre'):
            producto_nombre = items[0]['producto_nombre']
        else:
            r = c.execute(
                "SELECT producto FROM produccion_programada WHERE id=?",
                (produccion_id,)
            ).fetchone()
            producto_nombre = r[0] if r else ''

        if producto_nombre:
            def _read_meta():
                mr = c.execute("""
                    SELECT COALESCE(imagen_url,''), COALESCE(sku_principal,''),
                           COALESCE(descripcion_plain,''), COALESCE(precio_venta,0),
                           COALESCE(peso_g,0), COALESCE(imagenes_extra_json,'[]'),
                           COALESCE(shopify_handle,''), COALESCE(shopify_synced_at,'')
                    FROM formula_headers WHERE producto_nombre=?
                """, (producto_nombre,)).fetchone()
                return mr

            mr = _read_meta()

            # Auto-sync inmediato del producto si NUNCA se sincronizo (max 8s).
            # Asi el primer click al checklist trae la foto sin que Sebastian
            # tenga que hacer click manual.
            if mr and not mr[7]:  # shopify_synced_at vacio
                try:
                    from blueprints.inventario import _shopify_sync_producto
                    _shopify_sync_producto(conn, producto_nombre, timeout=8)
                    mr = _read_meta()  # re-leer despues del sync
                except Exception:
                    pass

            # Disparar sync masivo en background para los demas productos
            # (no bloquea esta respuesta — corre en thread separado)
            try:
                from blueprints.inventario import _sync_shopify_pendientes_background
                _sync_shopify_pendientes_background(max_edad_horas=24, max_productos=50)
            except Exception:
                pass

            if mr:
                import json as _json
                try:
                    imagenes_extra = _json.loads(mr[5] or '[]')
                except Exception:
                    imagenes_extra = []
                producto_meta = {
                    'imagen_url':       mr[0],
                    'sku':              mr[1],
                    'descripcion':      mr[2],
                    'precio':           float(mr[3] or 0),
                    'peso_g':           float(mr[4] or 0),
                    'imagenes_extra':   imagenes_extra,
                    'shopify_handle':   mr[6],
                    'shopify_synced_at': mr[7],
                }
    except Exception:
        pass

    return jsonify({
        'items': items,
        'totales_por_estado': totales,
        'total_items': total_items,
        'completados': completados,
        'porcentaje_listo': porcentaje,
        'producto_nombre': producto_nombre,
        # Compat con UI vieja
        'imagen_url': producto_meta.get('imagen_url', ''),
        # Metadata Shopify completa
        'producto_meta': producto_meta,
    })


@bp.route('/api/programacion/checklist/items/<int:item_id>', methods=['PATCH'])
def checklist_item_update(item_id):
    """Actualiza estado/proveedor/observaciones/etc de un item."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    fields = ['estado', 'proveedor', 'observaciones', 'responsable',
              'fecha_eta', 'fecha_recibido', 'fecha_solicitud',
              'solicitud_numero', 'oc_numero', 'cantidad_requerida']
    sets = []
    vals = []
    for f in fields:
        if f in d:
            sets.append(f'{f}=?')
            vals.append(d[f])
    if not sets:
        return jsonify({'error': 'Nada que actualizar'}), 400
    sets.append("actualizado_at=datetime('now', '-5 hours')")
    sets.append("actualizado_por=?")
    vals.append(session.get('compras_user', 'sistema'))
    vals.append(item_id)
    conn = get_db(); c = conn.cursor()
    c.execute(f"UPDATE produccion_checklist SET {', '.join(sets)} WHERE id=?", vals)
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/programacion/checklist/items/<int:item_id>/solicitar', methods=['POST'])
def checklist_item_solicitar(item_id):
    """Genera una SOL automatica para este item si requiere compra.

    Crea solicitud_compra con datos del item + actualiza estado a 'solicitado'.
    Si el item ya tiene solicitud_numero, no crea nueva.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    item = c.execute("""SELECT * FROM produccion_checklist WHERE id=?""", (item_id,)).fetchone()
    if not item:
        return jsonify({'error': 'Item no encontrado'}), 404
    item_dict = dict(zip([x[0] for x in c.description], item))

    if item_dict.get('solicitud_numero'):
        return jsonify({'ok': True, 'ya_solicitado': True,
                        'solicitud_numero': item_dict['solicitud_numero']})

    # ABASTECIMIENTO-FIX · 22-may-2026 · dedup cross-checklist (#7 audit 22-may)
    # · Antes: 5 producciones del mismo SKU → 5 SOLs separadas del mismo activo
    # · Ahora: si helper _pendiente_en_compras_g(codigo_mp) ya cubre cantidad
    #   requerida, NO crear SOL (link al ya existente queda como observación)
    codigo_mp_req = (item_dict.get('codigo_mp') or '').strip()
    cant_req = float(item_dict.get('cantidad_requerida') or 0)
    if codigo_mp_req and cant_req > 0:
        try:
            from blueprints.compras import _pendiente_en_compras_g
            en_cola = _pendiente_en_compras_g(c, codigo_mp_req)
            if en_cola >= cant_req:
                # Ya hay cola suficiente · marcar como solicitado sin duplicar
                # Buscar el SOL más reciente del mismo codigo_mp para link
                r_sol = c.execute(
                    """SELECT sc.numero FROM solicitudes_compra sc
                       JOIN solicitudes_compra_items sci ON sci.numero=sc.numero
                       WHERE sci.codigo_mp=?
                         AND sc.estado IN ('Pendiente','En revision','Aprobada')
                       ORDER BY sc.fecha DESC LIMIT 1""",
                    (codigo_mp_req,),
                ).fetchone()
                sol_link = r_sol[0] if r_sol else None
                if sol_link:
                    c.execute(
                        """UPDATE produccion_checklist
                           SET solicitud_numero=?, estado='solicitado'
                           WHERE id=?""",
                        (sol_link, item_id),
                    )
                    conn.commit()
                    return jsonify({
                        'ok': True, 'ya_solicitado': True,
                        'solicitud_numero': sol_link,
                        'dedup': True,
                        'en_cola_g': en_cola,
                        'mensaje': f'MP {codigo_mp_req} ya tiene {en_cola:.0f}g en cola (SOL {sol_link}) · vinculado sin duplicar',
                    })
        except Exception as _e:
            log.warning('checklist dedup _pendiente fallo: %s', _e)

    # Generar numero SOL
    from datetime import datetime as _dt
    year = _dt.now().year
    last = c.execute(
        "SELECT numero FROM solicitudes_compra WHERE numero LIKE ? ORDER BY numero DESC LIMIT 1",
        (f'SOL-{year}-%',)
    ).fetchone()
    seq = 1
    if last:
        try: seq = int(last[0].split('-')[-1]) + 1
        except: seq = 1
    sol_num = f'SOL-{year}-{seq:04d}'

    # Determinar categoria segun tipo
    # Fix #5 · 21-may-2026 · serigrafía/tampografía pasan a 'Material de
    # Empaque' (antes 'Servicios' → invisible en tab Planta de Catalina).
    # El servicio de decoración va sobre envases existentes pero Catalina
    # los gestiona junto al resto de empaque. Tipo='Servicio decoracion'
    # preserva la distinción interna.
    cat_map = {
        'mp': 'Materia Prima',
        'envase_primario': 'Material de Empaque',
        'envase_secundario': 'Material de Empaque',
        'tapa': 'Material de Empaque',
        'etiqueta_frontal': 'Material de Empaque',
        'etiqueta_posterior': 'Material de Empaque',
        'etiqueta_lateral': 'Material de Empaque',
        'caja_exterior': 'Material de Empaque',
        'serigrafia': 'Material de Empaque',
        'tampografia': 'Material de Empaque',
        'instructivo': 'Material de Empaque',
        'estuche': 'Material de Empaque',
    }
    categoria = cat_map.get(item_dict.get('item_tipo'), 'Material de Empaque')
    tipo_servicio_deco = item_dict.get('item_tipo') in ('serigrafia', 'tampografia')

    descripcion = (
        f"[Checklist Produccion] {item_dict.get('descripcion','')}"
        f" para {item_dict.get('producto_nombre','')}"
        f" (lote programado {item_dict.get('fecha_planeada','')})"
    )
    if item_dict.get('cantidad_requerida') and item_dict.get('item_tipo') == 'mp':
        descripcion += f" | Cantidad: {item_dict['cantidad_requerida']:.0f} {item_dict.get('unidad','g')}"

    user = session.get('compras_user', 'sistema')
    user_email = ''
    try:
        from config import USER_EMAILS as _UE
        user_email = _UE.get(user, '')
    except Exception:
        pass

    # Fix #8 · 21-may-2026 · urgencia dinámica según fecha planeada
    # (antes hardcoded 'Alta'). Si producción es mañana y MP falta → 'Critica'.
    urg_dinamica = 'Alta'
    try:
        from datetime import datetime as _dt, date as _date
        fp = item_dict.get('fecha_planeada')
        if fp:
            fpd = _dt.strptime(str(fp)[:10], '%Y-%m-%d').date()
            dias_falta = (fpd - _date.today()).days
            if dias_falta <= 1:
                urg_dinamica = 'Critica'
            elif dias_falta <= 3:
                urg_dinamica = 'Urgente'
            elif dias_falta <= 7:
                urg_dinamica = 'Alta'
            else:
                urg_dinamica = 'Normal'
    except Exception:
        pass
    tipo_sol = 'Servicio decoracion' if tipo_servicio_deco else 'Material'
    c.execute("""INSERT INTO solicitudes_compra
        (numero, fecha, estado, solicitante, email_solicitante, urgencia,
         observaciones, area, empresa, categoria, tipo)
        VALUES (?,date('now', '-5 hours'),'Pendiente',?,?,?,?,?,?,?,?)""",
        (sol_num, user, user_email, urg_dinamica, descripcion,
         'Produccion', 'Espagiria', categoria, tipo_sol))

    # Si es MP con codigo + cantidad: tambien insertar el item
    try:
        if item_dict.get('codigo_mp') and item_dict.get('cantidad_requerida'):
            c.execute("""INSERT INTO solicitudes_compra_items
                (numero, codigo_mp, nombre_mp, cantidad_g, justificacion)
                VALUES (?,?,?,?,?)""",
                (sol_num, item_dict['codigo_mp'],
                 item_dict.get('descripcion',''),
                 float(item_dict['cantidad_requerida']),
                 'Generado desde checklist pre-produccion'))
    except Exception:
        pass

    # Actualizar item del checklist
    c.execute("""UPDATE produccion_checklist
                 SET estado='solicitado', solicitud_numero=?,
                     fecha_solicitud=date('now', '-5 hours'),
                     actualizado_at=datetime('now', '-5 hours'), actualizado_por=?
                 WHERE id=?""",
              (sol_num, user, item_id))
    conn.commit()
    return jsonify({'ok': True, 'solicitud_numero': sol_num,
                    'mensaje': f'Solicitud {sol_num} creada para {item_dict.get("descripcion","")}'})


def _ensure_sync_log_table(conn):
    """Tabla minima para registrar el timestamp del ultimo sync de cada
    fuente externa (calendario, shopify, etc). Una fila por sync_type."""
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sync_log (
                sync_type    TEXT PRIMARY KEY,
                last_run_at  TEXT NOT NULL,
                last_count   INTEGER DEFAULT 0,
                last_error   TEXT
            )
        """)
    except Exception:
        pass


def _record_sync(conn, sync_type, count, error=None):
    """Registra timestamp del ultimo sync exitoso/fallido de un tipo."""
    _ensure_sync_log_table(conn)
    try:
        ts = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
        conn.execute("""
            INSERT INTO sync_log (sync_type, last_run_at, last_count, last_error)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(sync_type) DO UPDATE SET
                last_run_at = excluded.last_run_at,
                last_count  = excluded.last_count,
                last_error  = excluded.last_error
        """, (sync_type, ts, count, error))
        conn.commit()
    except Exception:
        pass


def _get_last_sync(conn, sync_type):
    """Lee timestamp y count del ultimo sync de un tipo. None si nunca."""
    _ensure_sync_log_table(conn)
    try:
        row = conn.execute(
            "SELECT last_run_at, last_count, last_error FROM sync_log WHERE sync_type=?",
            (sync_type,)
        ).fetchone()
        if not row:
            return None
        return {'last_run_at': row[0], 'last_count': row[1] or 0, 'last_error': row[2]}
    except Exception:
        return None


def _sync_calendar_a_produccion_programada(conn, days_ahead=90,
                                              force_mirror=False):
    """Sincroniza eventos de Google Calendar a la tabla produccion_programada.

    Idempotente: usa (producto, fecha_programada) como key para evitar
    duplicados. Marca origen='calendar' para distinguir de manuales.

    Sebastian (28-abr-2026): el checklist Pre-Produccion lee de
    produccion_programada, pero los eventos viven en el calendario
    animuslb.com. Antes habia que duplicar manualmente. Ahora se
    auto-sincroniza al cargar el checklist.

    Sebastian 7-may-2026: agregado force_mirror=True para hacer espejo
    DURO con Calendar · borra (HARD DELETE) cualquier producción del
    horizonte que NO esté en Calendar, sin importar origen. Solo
    respeta el guard de inicio_real_at / inventario_descontado_at.
    Útil cuando hay entries manuales viejas que sobrevivieron al sync
    bidireccional standard. Sin force_mirror, comportamiento legacy:
    solo cancela origen='calendar' orphans.

    Returns: count de filas insertadas en esta corrida.
    """
    import datetime as _dt
    import re as _re
    cal_error = None
    try:
        cal = _fetch_calendar_events(days_ahead=days_ahead)
    except Exception as _e:
        cal_error = str(_e)
        cal = {'events': [], 'error': cal_error}
    events = cal.get('events', [])
    # Sebastian 7-may-2026 (golden path 2): si NO hay events, antes hacíamos
    # early return y eso impedía que force_mirror borrase orfanos. Ahora:
    # · Si hay error de API → return (no destruir basado en data corrupta)
    # · Si la API respondió con 0 eventos legítimos → seguimos al cleanup
    #   pass (con keys_calendar=∅, TODO orfan será detectado correctamente).
    if cal_error or cal.get('error'):
        # FIX P0 audit Planta 24-may-2026 · antes el early-return NO
        # llamaba a _record_sync, así que la UI mostraba "OK · hace 2min"
        # cuando Calendar API estaba caída hace horas. Ahora registramos
        # el intento fallido para que el badge UI muestre el error real.
        _err_msg = cal_error or cal.get('error')
        _record_sync(conn, 'calendar', 0, error=_err_msg)
        return 0
    if not events and not force_mirror:
        # Sin events + sin espejo: nada que insertar, nada que borrar.
        # Aun así registramos el sync para que el admin sepa cuándo fue
        # la última ejecución exitosa (aunque vacía).
        _record_sync(conn, 'calendar', 0)
        return 0

    # Cargar mapa SKU → producto y formulas (para validar y obtener kg/lote)
    formulas = _get_formulas(conn)
    sku_to_prod = {}
    try:
        for row in conn.execute(
            "SELECT UPPER(sku), producto_nombre FROM sku_producto_map WHERE activo=1"
        ).fetchall():
            sku_to_prod[row[0]] = row[1]
    except Exception:
        pass

    # Filtrar tokens que NO son SKUs (mismo set que planificacion_estrategica)
    NOT_SKU = {
        'FAB','QC','CON','SIN','MICRO','ENVASADO','ACONDICIONAMIENTO',
        'FABRICACION','FABRICACIÓN','LANZAMIENTO','PRODUCCION','PRODUCCIÓN',
        'KG','MES','DIAS','DÍAS','ML','UDS','BATCH','FERNANDO',
        'MESA','LOTES','LOTE','MINI','MACRO','AND','THE','FOR',
        # Tokens de codigo de area que Alejandro escribe — no son SKUs
        'FAB1','FYE2','FYE3','ENV1','ENV2','PROD1','PROD2','PROD3','PROD4'
    }
    NON_FAB_KW = NON_FAB_KW_GLOBAL  # Sebastián 12-may-2026: unificado

    # Mapa codigo_corto_alejandro → codigo_real_db (areas_planta.codigo)
    # Sebastián 2-may-2026: "alejandro escribe [FAB1] o [FYE2] al inicio del
    # evento de Calendar y el sistema asigna sala automatica". Acepta tambien
    # los codigos reales (PROD1..PROD4) por si el equipo los usa directo.
    AREA_ALIAS = {
        'FAB1': 'PROD1', 'FYE2': 'PROD2', 'FYE3': 'PROD3', 'ENV2': 'PROD4',
        'ENV1': 'ENV1',
        'PROD1': 'PROD1', 'PROD2': 'PROD2', 'PROD3': 'PROD3', 'PROD4': 'PROD4',
    }
    # Mapa codigo_real_db → area_id (cargado una vez fuera del loop)
    codigo_a_areaid = {}
    try:
        for r in conn.execute(
            "SELECT id, codigo FROM areas_planta WHERE activo=1"
        ).fetchall():
            codigo_a_areaid[r[1]] = r[0]
    except Exception:
        pass

    def _skus(titulo):
        tokens = _re.findall(r'\b([A-Z][A-Z0-9]{1,}[A-Z0-9])\b', titulo.upper())
        return [t for t in tokens if t not in NOT_SKU]

    def _kg(titulo):
        m = _re.findall(r'~?(\d+(?:[,.]\d+)*)\s*kg', titulo, _re.IGNORECASE)
        if not m: return None
        try: return float(m[-1].replace(',', '.'))
        except Exception: return None

    def _area_from_titulo(titulo):
        """Extrae area_id del prefijo [CODIGO] del titulo del evento.

        Convención Alejandro (May 2026): el evento de Calendar empieza con
        un código entre corchetes, ej: '[FAB1] Gel Hidratante 50ml ~5kg' o
        '[FYE2] Blush Balm ~3kg'. Si lo encuentra, devuelve el area_id de
        areas_planta. Si no, None.
        """
        if not titulo:
            return None
        m = _re.match(r'\s*\[([A-Z0-9]{3,5})\]', titulo)
        if not m:
            return None
        codigo_corto = m.group(1).upper()
        codigo_real = AREA_ALIAS.get(codigo_corto)
        if not codigo_real:
            return None
        return codigo_a_areaid.get(codigo_real)

    today = _dt.date.today()
    inserted = 0
    # Sebastián 12-may-2026: contadores forenses para diagnosticar Calendar.
    # Antes era silent: si Calendar tenía 50 eventos y solo 10 entraban, no
    # había forma de saber cuántos se rechazaron y por qué. Ahora reporta:
    #   - eventos_no_fab: NON_FAB_KW match (reuniones, envasado, etc)
    #   - eventos_sin_sku: título sin token que matchee un SKU activo
    #   - eventos_sku_sin_formula: SKU mapeado pero sin fórmula en DB
    #   - eventos_fecha_pasada: fecha anterior a hoy
    #   - eventos_fecha_invalida: no se pudo parsear
    rechazados = {
        'no_fab': [], 'sin_sku': [], 'sku_sin_formula': [],
        'fecha_pasada': 0, 'fecha_invalida': 0,
    }
    for ev in events:
        titulo = ev.get('titulo', '')
        fecha_s = ev.get('fecha', '')
        gcal_id = ev.get('id', '') or ''
        if not fecha_s:
            rechazados['fecha_invalida'] += 1
            continue
        try:
            fecha = _dt.date.fromisoformat(fecha_s)
        except ValueError:
            rechazados['fecha_invalida'] += 1
            continue
        if fecha < today:
            rechazados['fecha_pasada'] += 1
            continue
        # Saltar eventos que NO son fabricacion (envasado/QC/etc consumen
        # nada de MPs crudas y no tienen sentido en el checklist).
        tlow = titulo.lower()
        if any(kw in tlow for kw in NON_FAB_KW):
            rechazados['no_fab'].append(titulo[:80])
            continue
        kg_evento = _kg(titulo)
        # Extraer area del prefijo [CODIGO] (Alejandro convention).
        area_id_titulo = _area_from_titulo(titulo)
        matched_any = False
        for sku in _skus(titulo):
            prod = sku_to_prod.get(sku)
            if not prod:
                continue
            if prod not in formulas:
                rechazados['sku_sin_formula'].append({'sku': sku, 'producto': prod, 'titulo': titulo[:80]})
                matched_any = True  # tenía SKU válido aunque sin fórmula
                continue
            matched_any = True
            lote_kg = formulas[prod].get('lote_size_kg', 0) or 0
            # Sebastián 12-may-2026: si hay kg_evento del título, usar 1 lote
            # con cantidad_kg=kg_evento (NO dividir por lote_size_kg para
            # calcular múltiples lotes — Calendar representa UNA producción).
            if kg_evento and kg_evento > 0:
                lotes = 1
                cantidad_kg_calc = kg_evento
            else:
                lotes = 1
                cantidad_kg_calc = lote_kg if lote_kg > 0 else 0
            # INSERT idempotente prioritario por gcal_event_id, fallback (producto, fecha)
            try:
                exists = None
                if gcal_id:
                    exists = conn.execute(
                        "SELECT id, COALESCE(cantidad_kg,0), area_id "
                        "FROM produccion_programada "
                        "WHERE gcal_event_id=? LIMIT 1",
                        (gcal_id,)
                    ).fetchone()
                if not exists:
                    exists = conn.execute(
                        "SELECT id, COALESCE(cantidad_kg,0), area_id "
                        "FROM produccion_programada "
                        "WHERE producto=? AND fecha_programada=? LIMIT 1",
                        (prod, fecha_s)
                    ).fetchone()
                if not exists:
                    # origen='calendar' para que planificacion_estrategica las
                    # filtre y no duplique con su lectura directa del calendar.
                    # area_id = del prefijo [CODIGO] si Alejandro lo puso, sino NULL.
                    cur_ins = conn.execute(
                        """INSERT INTO produccion_programada
                           (producto, fecha_programada, lotes, cantidad_kg,
                            estado, observaciones, origen, area_id, gcal_event_id)
                           VALUES (?, ?, ?, ?, 'programado', ?, 'calendar', ?, ?)""",
                        (prod, fecha_s, lotes, cantidad_kg_calc,
                         f'[auto-sync calendar] {titulo[:200]}', area_id_titulo,
                         gcal_id)
                    )
                    inserted += 1
                    # Sebastián 1-may-2026: "todo automático". Auto-asignar
                    # área + operarios INMEDIATAMENTE al insertar (no esperar
                    # cron 06:30). Si falla, sigue → cron lo intentará después.
                    # Si Alejandro ya puso [CODIGO] en titulo, NO sobrescribir
                    # con auto-asignador: respeta su decision.
                    try:
                        new_id = cur_ins.lastrowid
                        if (new_id and cantidad_kg_calc and cantidad_kg_calc > 0
                                and area_id_titulo is None):
                            _auto_asignar_produccion(conn.cursor(), new_id, 'auto-sync-calendar')
                    except Exception:
                        pass
                else:
                    # P0-9 23-may-PM · auditoría agente · los 3 backfills
                    # UPDATE NO excluían origen Fijo · si Alejandro ponía
                    # [CODIGO] en un evento Calendar que matcheaba (producto,
                    # fecha) con un Fijo, el sync MUTABA el Fijo. Violación
                    # de INV-6 Fijo intocable por procesos automáticos.
                    # Filtro agregado a los 3 UPDATEs.
                    _NO_FIJO = ("AND COALESCE(origen,'') NOT IN "
                                 "('eos_plan','eos_b2b','eos_retroactivo')")
                    if (exists[1] or 0) <= 0 and cantidad_kg_calc > 0:
                        conn.execute(
                            f"UPDATE produccion_programada SET cantidad_kg=? "
                            f"WHERE id=? {_NO_FIJO}",
                            (cantidad_kg_calc, exists[0])
                        )
                    if area_id_titulo is not None and (exists[2] is None):
                        conn.execute(
                            f"UPDATE produccion_programada SET area_id=? "
                            f"WHERE id=? {_NO_FIJO}",
                            (area_id_titulo, exists[0])
                        )
                    if gcal_id:
                        try:
                            conn.execute(
                                f"UPDATE produccion_programada SET gcal_event_id=? "
                                f"WHERE id=? AND COALESCE(gcal_event_id,'') = '' {_NO_FIJO}",
                                (gcal_id, exists[0])
                            )
                        except Exception:
                            pass
            except Exception:
                continue
            break  # un evento → un producto match
        if not matched_any:
            rechazados['sin_sku'].append(titulo[:80])
    if inserted:
        conn.commit()

    # ── Sync BIDIRECCIONAL ─────────────────────────────────────────────
    # Marcar como 'cancelado' las filas con origen='calendar' cuya
    # (producto, fecha_programada) YA NO existe en el calendar — es decir,
    # Sebastian las borro o movio en Google Calendar y aqui quedaron de
    # huerfanas. Si NO hacemos esto, el horizonte del checklist mostrara
    # producciones fantasma que ya no estan programadas.
    # Solo toca origen='calendar' — las manuales (origen NULL/manual) NO
    # se tocan jamas porque no tienen contraparte en el calendar.
    #
    # Audit zero-error 2-may-2026 (CERO SESGO):
    # GUARD CRÍTICO · NO cancelar producciones que ya iniciaron o ya
    # descontaron inventario. Si Sebastián borra un evento del Calendar
    # MIENTRAS la producción está en curso, hacerla 'cancelado' aquí
    # corromperia el estado y crearía drift de inventario. En ese caso,
    # hay que dejarla activa y registrar la acción en audit_log para que
    # alguien revise manualmente (probablemente fue un error en Calendar).
    archived = 0
    skipped_in_progress = []
    try:
        # Set de (producto, fecha) que SI existen actualmente en calendar
        keys_calendar = set()
        for ev in events:
            tlow = ev.get('titulo', '').lower()
            if any(kw in tlow for kw in NON_FAB_KW):
                continue
            fecha_s = ev.get('fecha', '')
            if not fecha_s:
                continue
            for sku in _skus(ev.get('titulo', '')):
                prod = sku_to_prod.get(sku)
                if prod and prod in formulas:
                    keys_calendar.add((prod, fecha_s))
                    break
        # Sebastian 7-may-2026 · modo espejo:
        #   force_mirror=False (default): solo origen='calendar' (legacy)
        #   force_mirror=True: cualquier origen · ESPEJO DURO con Calendar
        # IMPORTANTE: capturar inicio_real_at + inventario_descontado_at para
        # decidir si es seguro cancelar (no corromper inventario en curso).
        if force_mirror:
            # Sebastián 19-may-2026: el espejo duro NUNCA borra lo que el
            # usuario fijó (eos_plan) ni pedidos B2B ni históricos. Solo
            # espeja lo que es sugerencia (canónico / calendar / manual).
            origen_filter = ("AND COALESCE(origen,'') "
                             "NOT IN ('eos_plan','eos_b2b','eos_retroactivo')")
        else:
            origen_filter = "AND origen='calendar'"
        # Sebastián 12-may-2026: ventana ampliada de -1d a -14d.
        # Antes el sync solo limpiaba fechas >=hoy-1, así que las atrasadas
        # zombies (Alejandro movió/borró en Calendar pero quedaron en BD)
        # del pasado nunca se cancelaban → panel mostraba "Mar 5 atrasada"
        # cuando Calendar nunca tuvo nada Mar 5. Ahora se cancelan también
        # las de hasta 14 días atrás SI no iniciaron Y no descontaron inv.
        candidatos = conn.execute(
            f"""SELECT id, producto, fecha_programada,
                       COALESCE(inicio_real_at,'') as inicio,
                       COALESCE(inventario_descontado_at,'') as descontado,
                       COALESCE(origen,'manual') as origen_val
               FROM produccion_programada
               WHERE 1=1 {origen_filter}
                 AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
                 AND fecha_programada >= date('now', '-5 hours', '-14 days')"""
        ).fetchall()
        huerfanos = []
        for r in candidatos:
            id_, prod, fecha, inicio, descontado, origen_val = r
            if (prod, fecha) in keys_calendar:
                continue  # sigue en calendar, no es huérfano
            # GUARD: si ya inició o ya descontó inventario, NO tocar
            if inicio or descontado:
                skipped_in_progress.append({
                    'id': id_, 'producto': prod, 'fecha': fecha,
                    'origen': origen_val,
                    'inicio_real_at': inicio or None,
                    'inventario_descontado_at': descontado or None,
                })
                continue
            huerfanos.append(id_)
        if huerfanos:
            placeholders = ','.join(['?'] * len(huerfanos))
            if force_mirror:
                # Modo espejo · HARD DELETE para que la app refleje exactamente
                # Calendar (sin entries fantasma cancelados ocupando memoria).
                conn.execute(
                    f"DELETE FROM produccion_programada WHERE id IN ({placeholders}) "
                    f"AND COALESCE(inicio_real_at,'')='' "
                    f"AND COALESCE(inventario_descontado_at,'')='' "
                    f"AND COALESCE(origen,'') NOT IN ('eos_plan','eos_b2b','eos_retroactivo')",
                    huerfanos,
                )
                accion_audit = 'AUTO_BORRAR_PRODUCCION_ESPEJO'
                detalle_audit = 'force_mirror: sync borró producción no-en-calendar'
            else:
                # Modo legacy · solo cancela (preserva historial)
                conn.execute(
                    f"UPDATE produccion_programada SET estado='cancelado', "
                    f"observaciones=COALESCE(observaciones,'') || ' [auto-cancelado: ya no esta en calendar]' "
                    f"WHERE id IN ({placeholders})",
                    huerfanos,
                )
                accion_audit = 'AUTO_CANCELAR_PRODUCCION'
                detalle_audit = 'evento ya no existe en calendar'
            archived = len(huerfanos)
            # Audit log · cada acción auto debe quedar trazada
            try:
                from audit_helpers import audit_log as _audit_log
                cur = conn.cursor()
                for prod_id in huerfanos:
                    _audit_log(cur, usuario='sistema_calendar_sync',
                               accion=accion_audit,
                               tabla='produccion_programada',
                               registro_id=prod_id,
                               despues={'razon': detalle_audit,
                                         'force_mirror': force_mirror},
                               detalle=f"Sync calendar id={prod_id} · {detalle_audit}")
            except Exception:
                pass
            conn.commit()
        # Audit log para producciones que se SALTARON cancelar (en curso)
        if skipped_in_progress:
            try:
                from audit_helpers import audit_log as _audit_log
                cur = conn.cursor()
                for sk in skipped_in_progress:
                    _audit_log(cur, usuario='sistema_calendar_sync',
                               accion='SYNC_CALENDAR_SKIP_EN_CURSO',
                               tabla='produccion_programada',
                               registro_id=sk['id'],
                               detalle=(f"Producción {sk['producto']} ({sk['fecha']}) "
                                         f"removida del calendar pero está EN CURSO "
                                         f"(inicio={sk['inicio_real_at']}, "
                                         f"descontado={sk['inventario_descontado_at']}). "
                                         f"NO se canceló · revisar manualmente."))
                conn.commit()
            except Exception:
                pass
    except Exception:
        archived = 0
        skipped_in_progress = []

    # Registrar timestamp del sync (independiente de si hubo nuevas)
    _record_sync(conn, 'calendar', inserted + archived, error=cal.get('error'))

    # Sebastián 12-may-2026: log forense de rechazos (visibles vía logs Render).
    # Si hay muchos no_fab o sin_sku, es señal de que Calendar tiene títulos
    # raros que no matchean SKUs activos · útil para limpiar.
    n_no_fab = len(rechazados.get('no_fab', []))
    n_sin_sku = len(rechazados.get('sin_sku', []))
    n_sin_formula = len(rechazados.get('sku_sin_formula', []))
    if inserted or n_no_fab or n_sin_sku or n_sin_formula:
        log.info(
            'sync_calendar · inserted=%d archived=%d skip_in_progress=%d '
            'rechazos: no_fab=%d sin_sku=%d sku_sin_formula=%d '
            'fecha_pasada=%d fecha_invalida=%d',
            inserted, archived, len(skipped_in_progress),
            n_no_fab, n_sin_sku, n_sin_formula,
            rechazados.get('fecha_pasada', 0),
            rechazados.get('fecha_invalida', 0),
        )

    # Sebastián 12-may-2026: retornar dict con detalles para que callers
    # puedan exponer 'cosas raras' al usuario. Compat: si caller solo usa
    # como int, hace .get('inserted', 0) o similar.
    return {
        'inserted': inserted,
        'archived': archived,
        'skipped_in_progress': skipped_in_progress,
        'rechazos': rechazados,
        'events_count': len(events),
        'error': cal.get('error'),
    }


def _start_calendar_sync_background_loop():
    """Arranca un thread daemon que cada N minutos sincroniza el calendario
    sin requerir que alguien tenga la pantalla abierta. Idempotente: solo
    arranca una instancia por proceso. Se reinicia al primer request del
    blueprint si Render hace cold-start.

    Frecuencia configurable via env var CALENDAR_SYNC_INTERVAL_MIN
    (default 10 min). Si <=0, queda desactivado.

    Decision Sebastian 2026-04-29: el checklist debe estar fresco aunque
    nadie lo este mirando — Catalina entra en la mañana y la cola ya
    refleja lo que se agrego al calendar.
    """
    import threading, os, time as _t
    if getattr(_start_calendar_sync_background_loop, '_running', False):
        return
    # Sebastián 14-may-2026: "solo aparezca lo que construí contigo · es
    # la realidad". Calendar legacy se reemplaza por canónicos eos_canonico
    # configurables en /admin/configurar-canonicos. Default desactivado.
    # Si se necesita reactivar, setear env CALENDAR_SYNC_INTERVAL_MIN > 0
    # explícitamente en Render.
    try:
        interval_min = int(os.environ.get('CALENDAR_SYNC_INTERVAL_MIN', '0'))
    except ValueError:
        interval_min = 0
    if interval_min <= 0:
        return  # desactivado por default (Sebastián 14-may-2026)
    _start_calendar_sync_background_loop._running = True

    def _worker():
        from config import DB_PATH
        import sqlite3
        # Pequeño delay inicial para no chocar con el startup del proceso
        _t.sleep(20)
        while True:
            try:
                local_conn = db_connect(timeout=30)
                try:
                    _sync_calendar_a_produccion_programada(local_conn, days_ahead=120)
                finally:
                    local_conn.close()
            except Exception:
                pass  # el loop nunca debe morir por una falla puntual
            _t.sleep(max(60, interval_min * 60))

    t = threading.Thread(target=_worker, daemon=True, name='calendar-sync-loop')
    t.start()


# Arrancar el loop al importar el blueprint (una vez por proceso)
try:
    _start_calendar_sync_background_loop()
except Exception:
    pass


@bp.route('/api/programacion/checklist/sync-calendar', methods=['POST'])
def checklist_sync_calendar_endpoint():
    """Endpoint manual para forzar sincronizacion calendario → produccion_programada.

    Tambien se llama automaticamente al cargar /resumen-calendario y por
    un thread background cada N min (CALENDAR_SYNC_INTERVAL_MIN).

    Querystring:
      ?dias=90 · ventana hacia adelante
      ?force_mirror=true · modo espejo · HARD DELETE de cualquier
        producción del horizonte que NO esté en Calendar (cualquier
        origen, NO solo 'calendar') · respeta guard inicio/descontado.
        Sin este flag: comportamiento legacy (solo cancela calendar).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    days = int(request.args.get('dias', 90))
    force_mirror = (request.args.get('force_mirror', '').lower()
                    in ('true', '1', 'yes'))
    # P0 audit 26-may · INV-3 dice "force_mirror solo desde el botón
    # Re-sync Calendar (admin)" pero el endpoint solo checaba sesión ·
    # cualquier user con sesión podía pasar ?force_mirror=true y disparar
    # HARD DELETE bulk del horizonte.
    if force_mirror:
        _u = session.get('compras_user', '')
        try:
            from config import ADMIN_USERS as _AU
        except Exception:
            _AU = {'sebastian', 'alejandro'}
        if (_u or '').lower() not in {x.lower() for x in _AU}:
            return jsonify({'error': 'force_mirror requiere admin (HARD DELETE bulk)'}), 403
    conn = get_db()
    result = _sync_calendar_a_produccion_programada(
        conn, days_ahead=days, force_mirror=force_mirror,
    )
    # Backward compat: result puede ser dict (nuevo) o int (legacy en algún path).
    if isinstance(result, dict):
        inserted = result.get('inserted', 0)
        archived = result.get('archived', 0)
        rechazos = result.get('rechazos', {})
        events_count = result.get('events_count', 0)
    else:
        inserted = int(result or 0)
        archived = 0
        rechazos = {}
        events_count = 0
    last = _get_last_sync(conn, 'calendar')
    msg = f'{inserted} producciones nuevas importadas del calendario'
    if not inserted:
        msg = 'El calendario ya estaba sincronizado'
    if force_mirror:
        msg += ' · modo espejo (orfanos eliminados)'
    return jsonify({
        'ok': True,
        'producciones_creadas': inserted,
        'archivadas': archived,
        'events_total': events_count,
        'rechazos': {
            'no_fabricacion': len(rechazos.get('no_fab', [])),
            'sin_sku_mapeado': len(rechazos.get('sin_sku', [])),
            'sku_sin_formula': len(rechazos.get('sku_sin_formula', [])),
            'fecha_pasada': rechazos.get('fecha_pasada', 0),
            'fecha_invalida': rechazos.get('fecha_invalida', 0),
        },
        'rechazos_detalle': {
            'no_fabricacion_titulos': rechazos.get('no_fab', [])[:20],
            'sin_sku_titulos': rechazos.get('sin_sku', [])[:20],
            'sku_sin_formula_items': rechazos.get('sku_sin_formula', [])[:20],
        },
        'force_mirror': force_mirror,
        'last_run_at': last['last_run_at'] if last else None,
        'mensaje': msg,
    })


@bp.route('/api/programacion/checklist/resumen-calendario')
def checklist_resumen_calendario():
    """Vista panoramica: para cada produccion programada en proximos N dias,
    devuelve % de completitud del checklist + dias faltantes + items criticos
    pendientes.

    Auto-sincroniza eventos del calendario antes de listar para que las
    producciones nuevas del calendar aparezcan sin que el usuario tenga que
    hacer trigger manual. Usado por Planta + Gerencia.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    horizonte = int(request.args.get('dias', 60))
    conn = get_db(); c = conn.cursor()

    # Auto-sync calendario → produccion_programada (idempotente, falla silenciosa)
    sync_count = 0
    try:
        _sync_res = _sync_calendar_a_produccion_programada(conn, days_ahead=max(horizonte, 90))
        if isinstance(_sync_res, dict):
            sync_count = _sync_res.get('inserted', 0)
        else:
            sync_count = int(_sync_res or 0)
    except Exception:
        pass

    # Auto-sync masivo Shopify de productos pendientes en BACKGROUND.
    # No bloquea la respuesta — al refresh el usuario ya tiene fotos.
    try:
        from blueprints.inventario import _sync_shopify_pendientes_background
        _sync_shopify_pendientes_background(max_edad_horas=24, max_productos=50)
    except Exception:
        pass

    # Filtro legacy aplicado a cada subquery de conteo (etiquetas frontal/posterior/lateral
    # ya no cuentan porque la decoracion del envase las cubre). Se construye dinamicamente
    # para que cambiar TIPOS_LEGACY_OCULTOS arriba se propague aqui.
    if TIPOS_LEGACY_OCULTOS:
        legacy_sql = " AND item_tipo NOT IN (" + ",".join(["'" + t.replace("'", "''") + "'" for t in TIPOS_LEGACY_OCULTOS]) + ")"
    else:
        legacy_sql = ""
    try:
        rows = c.execute(f"""
            SELECT pp.id,
                   pp.producto                                       as producto_nombre,
                   pp.fecha_programada                               as fecha_planeada,
                   COALESCE(NULLIF(pp.cantidad_kg, 0),
                            COALESCE(pp.lotes, 1) * COALESCE(fh.lote_size_kg, 0)) as kg,
                   pp.estado                                         as estado_prod,
                   COALESCE(pp.origen, 'manual')                     as origen,
                   COALESCE(pp.inventario_descontado_at, '')         as descontado_at,
                   (julianday(pp.fecha_programada) - julianday('now')) as dias_faltan,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}) as total_items,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}
                       AND estado IN ('verificado_ok','recibido','listo')) as completados,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}
                       AND estado='pendiente') as pendientes,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}
                       AND estado='solicitado') as solicitados,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}
                       AND estado='en_transito') as en_transito,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}
                       AND estado='recibido') as recibidos,
                   (SELECT COUNT(*) FROM produccion_checklist WHERE produccion_id=pp.id{legacy_sql}
                       AND estado='no_aplica') as no_aplica
            FROM produccion_programada pp
            LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado')
              AND pp.fecha_programada >= date('now', '-5 hours', '-30 day')
              AND pp.fecha_programada <= date('now', '-5 hours', '+' || ? || ' day')
            ORDER BY pp.fecha_programada ASC
        """, (horizonte,)).fetchall()
        cols = [x[0] for x in c.description]
        out = []
        for r in rows:
            d = dict(zip(cols, r))
            total = d['total_items'] or 0
            comp = d['completados'] or 0
            d['porcentaje'] = round(comp / total * 100, 1) if total > 0 else 0
            d['semaforo'] = (
                'verde'    if d['porcentaje'] >= 90 else
                'amarillo' if d['porcentaje'] >= 50 else
                'rojo'
            )
            d['dias_faltan'] = int(d['dias_faltan']) if d['dias_faltan'] is not None else 0
            out.append(d)
        last_sync = _get_last_sync(conn, 'calendar')
        return jsonify({
            'producciones': out,
            'horizonte_dias': horizonte,
            'sync_calendario': {
                'producciones_nuevas': sync_count,
                'last_run_at': last_sync['last_run_at'] if last_sync else None,
                'last_error': last_sync['last_error'] if last_sync else None,
            },
        })
    except Exception as e:
        return jsonify({'error': str(e), 'producciones': []})


@bp.route('/api/programacion/disponibilidad-mp/<path:codigo_mp>')
def disponibilidad_mp_endpoint(codigo_mp):
    """Panorama completo de disponibilidad de un MP:
       stock + en_transito + solicitado - demanda_horizonte = neta.
    Querystring: ?dias=60 (default)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    dias = int(request.args.get('dias', 60))
    fecha_h = (datetime.now() + timedelta(days=dias)).strftime('%Y-%m-%d')
    conn = get_db(); c = conn.cursor()
    disp = _calcular_disponibilidad_mp(c, codigo_mp, fecha_horizonte=fecha_h)
    # Detalle de OCs en transito + SOLs pendientes
    try:
        ocs = c.execute("""
            SELECT oc.numero_oc, oc.estado, oc.proveedor,
                   (i.cantidad_g - COALESCE(i.cantidad_recibida_g,0)) as faltante_g,
                   oc.fecha_entrega_est
            FROM ordenes_compra_items i
            JOIN ordenes_compra oc ON oc.numero_oc=i.numero_oc
            WHERE i.codigo_mp=?
              AND oc.estado IN ('Borrador','Pendiente','Revisada','Aprobada',
                                'Autorizada','Parcial','Pagada')
              AND (i.cantidad_g - COALESCE(i.cantidad_recibida_g,0)) > 0
            ORDER BY oc.fecha DESC
        """, (codigo_mp,)).fetchall()
        disp['ocs_en_transito'] = [
            {'numero_oc': r[0], 'estado': r[1], 'proveedor': r[2],
             'cantidad_g': r[3], 'eta': r[4]} for r in ocs
        ]
    except Exception:
        disp['ocs_en_transito'] = []

    try:
        sols = c.execute("""
            SELECT s.numero, s.estado, s.fecha, si.cantidad_g, si.justificacion
            FROM solicitudes_compra_items si
            JOIN solicitudes_compra s ON s.numero=si.numero
            WHERE si.codigo_mp=?
              AND s.estado IN ('Pendiente','En revision','Aprobada')
              AND COALESCE(s.numero_oc,'')=''
            ORDER BY s.fecha DESC
        """, (codigo_mp,)).fetchall()
        disp['solicitudes_pendientes'] = [
            {'numero': r[0], 'estado': r[1], 'fecha': r[2],
             'cantidad_g': r[3], 'justificacion': r[4]} for r in sols
        ]
    except Exception:
        disp['solicitudes_pendientes'] = []

    # Producciones que demandan este MP
    try:
        prods = c.execute("""
            SELECT pp.id,
                   pp.producto                                          as producto_nombre,
                   pp.fecha_programada                                  as fecha_planeada,
                   COALESCE(pp.lotes, 1) * COALESCE(fh.lote_size_kg, 0) as kg,
                   fi.porcentaje,
                   ROUND(COALESCE(fi.cantidad_g_por_lote,0) * COALESCE(pp.lotes, 1)) as req_g
            FROM produccion_programada pp
            JOIN formula_items fi ON fi.producto_nombre = pp.producto
            LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE fi.material_id = ?
              AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado')
              AND pp.fecha_programada >= date('now', '-5 hours', '-1 day')
              AND pp.fecha_programada <= ?
            ORDER BY pp.fecha_programada ASC
        """, (codigo_mp, fecha_h)).fetchall()
        disp['producciones_que_lo_usan'] = [
            {'produccion_id': r[0], 'producto': r[1], 'fecha': r[2],
             'kg': r[3], 'porcentaje': r[4], 'requerido_g': r[5]}
            for r in prods
        ]
    except Exception:
        disp['producciones_que_lo_usan'] = []

    disp['codigo_mp'] = codigo_mp
    disp['horizonte_dias'] = dias
    disp['fecha_horizonte'] = fecha_h
    return jsonify(disp)


@bp.route('/api/programacion/checklist/backfill', methods=['POST'])
def checklist_backfill():
    """Genera checklists para producciones de los ultimos 30 dias en
    adelante que aun no tienen items creados.

    Solo admin. Idempotente.
    """
    if 'compras_user' not in session or session.get('compras_user') not in ADMIN_USERS:
        return jsonify({'error': 'Solo admins'}), 403

    # Wrapper externo: captura CUALQUIER excepcion (incluyendo get_db() o
    # imports rotos) para que devolvamos el traceback al frontend en JSON
    # en lugar de caer al @app.errorhandler(Exception) global de api/index.py
    # que solo dice "Error interno del servidor" sin pista del problema real.
    import traceback as _tb
    try:
        conn = get_db(); c = conn.cursor()
        rows = []
        try:
            rows = c.execute("""
                SELECT pp.id,
                       pp.producto                                          as producto_nombre,
                       pp.fecha_programada                                  as fecha_planeada,
                       COALESCE(NULLIF(pp.cantidad_kg, 0),
                                COALESCE(pp.lotes, 1) * COALESCE(fh.lote_size_kg, 0)) as kg
                FROM produccion_programada pp
                LEFT JOIN formula_headers fh
                       ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
                WHERE LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado','completado')
                  AND pp.fecha_programada >= date('now', '-5 hours', '-30 day')
                  AND NOT EXISTS (SELECT 1 FROM produccion_checklist
                                  WHERE produccion_id=pp.id)
            """).fetchall()
        except Exception as e:
            return jsonify({
                'error': f'Falla al listar producciones pendientes de checklist: {e}',
                'fase': 'select_pendientes',
                'traceback': _tb.format_exc()[-1500:],
            }), 500

        # Procesar cada produccion en su propio try/except — si una falla, las
        # demas siguen. Asi el backfill no se aborta por un solo producto sin
        # formula completa.
        total_creados = 0
        producciones_procesadas = 0
        fallas = []
        for r in rows:
            try:
                items = _generar_checklist_produccion(
                    c, r[0], r[1], r[2], float(r[3] or 0),
                    generar_mps=True,
                    usuario=session.get('compras_user', 'sistema'))
                total_creados += items
                producciones_procesadas += 1
            except Exception as e:
                fallas.append({
                    'produccion_id': r[0],
                    'producto': r[1],
                    'fecha': r[2],
                    'kg': r[3],
                    'error': str(e),
                    'traceback': _tb.format_exc()[-800:],
                })
        try:
            conn.commit()
        except Exception:
            pass
        payload = {
            'ok': len(fallas) == 0,
            'producciones_procesadas': producciones_procesadas,
            'items_creados': total_creados,
            'producciones_pendientes': len(rows),
            'fallas': fallas,
            'mensaje': (f'{producciones_procesadas} producciones recibieron checklist '
                        f'({total_creados} items totales)') if not fallas else (
                        f'{producciones_procesadas}/{len(rows)} OK · '
                        f'{len(fallas)} fallaron — ver "fallas" en la respuesta'),
        }
        # Devolver 200 incluso con fallas parciales (el frontend ve el detalle).
        return jsonify(payload)
    except Exception as e:
        return jsonify({
            'error': f'Falla inesperada en backfill (fuera del bucle): {e}',
            'fase': 'wrapper_externo',
            'tipo': type(e).__name__,
            'traceback': _tb.format_exc()[-2000:],
        }), 500


# ════════════════════════════════════════════════════════════════════════
# CHECKLIST EDITABLE + SOLICITUDES PRODUCCION + TAREAS OPERATIVAS
# Sebastian (29-abr-2026): cada item del checklist se hace editable, con
# dropdown de maestro_mee, cantidad calculada automaticamente, y boton
# "Solicitar" que crea entrada en cola de Catalina. Catalina decide:
# inventario / OC / serigrafia / tampografia (genera tarea operativa).
# ════════════════════════════════════════════════════════════════════════

@bp.route('/api/checklist/mee-options', methods=['GET'])
def checklist_mee_options():
    """Devuelve materiales de maestro_mee filtrados por tipo del item.

    Querystring: ?tipo=envase_primario|tapa|etiqueta_frontal|etc
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    tipo = request.args.get('tipo', '')
    # Mapeo tipo_item -> categoria/keywords en maestro_mee
    keyword_map = {
        'envase_primario':    ['envase', 'frasco', 'bote', 'tubo', 'pote'],
        'envase_secundario':  ['caja', 'estuche', 'secundario'],
        'tapa':               ['tapa', 'dosificador', 'pump', 'gotero', 'spray'],
        'etiqueta_frontal':   ['etiqueta', 'label'],
        'etiqueta_posterior': ['etiqueta', 'label'],
        'etiqueta_lateral':   ['etiqueta', 'label'],
        'caja_exterior':      ['caja', 'corrugada', 'exterior'],
        'instructivo':        ['instructivo', 'inserto'],
    }
    keywords = keyword_map.get(tipo, [])
    conn = get_db(); c = conn.cursor()
    if keywords:
        # OR de LIKE por cada keyword
        like_clauses = " OR ".join(["LOWER(descripcion) LIKE ?" for _ in keywords])
        params = [f"%{k}%" for k in keywords]
        rows = c.execute(f"""
            SELECT codigo, descripcion, COALESCE(stock_actual,0), unidad,
                   COALESCE(proveedor,''), COALESCE(categoria,'')
            FROM maestro_mee
            WHERE COALESCE(estado,'Activo')='Activo'
              AND ({like_clauses})
            ORDER BY descripcion
        """, params).fetchall()
    else:
        rows = c.execute("""
            SELECT codigo, descripcion, COALESCE(stock_actual,0), unidad,
                   COALESCE(proveedor,''), COALESCE(categoria,'')
            FROM maestro_mee
            WHERE COALESCE(estado,'Activo')='Activo'
            ORDER BY descripcion
            LIMIT 200
        """).fetchall()
    return jsonify({
        'tipo': tipo,
        'options': [
            {'codigo': r[0], 'descripcion': r[1], 'stock': float(r[2] or 0),
             'unidad': r[3] or 'und', 'proveedor': r[4], 'categoria': r[5]}
            for r in rows
        ]
    })


@bp.route('/api/programacion/checklist/items/<int:item_id>/asignar-mee', methods=['POST'])
def checklist_item_asignar_mee(item_id):
    """Guarda el material MEE elegido para un item + cantidad calculada.

    Body: {mee_codigo, cantidad_unidades, decoracion_tipo (opcional)}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    mee_codigo = (d.get('mee_codigo') or '').strip()
    cantidad = float(d.get('cantidad_unidades') or 0)
    decoracion = (d.get('decoracion_tipo') or '').strip()
    descripcion_actualizada = (d.get('descripcion') or '').strip()

    conn = get_db(); c = conn.cursor()
    # Si MEE existe, traer su descripcion para poblar el item
    desc_mee = ''
    if mee_codigo:
        rm = c.execute("SELECT descripcion FROM maestro_mee WHERE codigo=?", (mee_codigo,)).fetchone()
        desc_mee = rm[0] if rm else ''

    sets = ['mee_codigo_asignado=?', 'cantidad_unidades=?', 'actualizado_at=datetime(\'now\')']
    params = [mee_codigo, cantidad]
    if decoracion:
        sets.append('decoracion_tipo=?'); params.append(decoracion)
    if desc_mee:
        sets.append('codigo_mp=?'); params.append(mee_codigo)
    if descripcion_actualizada:
        sets.append('descripcion=?'); params.append(descripcion_actualizada)
    elif desc_mee:
        sets.append('descripcion=?'); params.append(desc_mee)
    params.append(item_id)
    c.execute(f"UPDATE produccion_checklist SET {', '.join(sets)} WHERE id=?", params)
    if c.rowcount == 0:
        return jsonify({'error': 'Item no encontrado'}), 404
    conn.commit()
    return jsonify({'ok': True, 'item_id': item_id, 'mee_codigo': mee_codigo,
                    'cantidad_unidades': cantidad, 'descripcion': desc_mee or descripcion_actualizada})


@bp.route('/api/programacion/checklist/items/<int:item_id>/solicitar-produccion', methods=['POST'])
def checklist_item_solicitar_produccion(item_id):
    """Genera una entrada en solicitudes_compra_anticipada (cola de Catalina) para el
    item del checklist. Catalina luego decide ruta (inventario/OC/serigrafia).

    Body opcional: {fecha_objetivo, observaciones}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    item = c.execute("""
        SELECT id, produccion_id, producto_nombre, item_tipo, descripcion,
               cantidad_requerida, cantidad_unidades, mee_codigo_asignado,
               decoracion_tipo, fecha_planeada, codigo_mp
        FROM produccion_checklist WHERE id=?
    """, (item_id,)).fetchone()
    if not item:
        return jsonify({'error': 'Item no encontrado'}), 404

    cantidad = float(item[6] or 0) or float(item[5] or 0)
    fecha_obj = (d.get('fecha_objetivo') or item[9] or '').strip()
    observ = (d.get('observaciones') or '').strip()

    # Si ya tiene una solicitud_produccion_id activa, devolverla (idempotente)
    existing = c.execute(
        "SELECT id FROM solicitudes_compra_anticipada WHERE checklist_item_id=? AND estado IN ('pendiente','decidida')",
        (item_id,)
    ).fetchone()
    if existing:
        return jsonify({'ok': True, 'solicitud_id': existing[0], 'ya_existia': True})

    cur = c.execute("""
        INSERT INTO solicitudes_compra_anticipada
          (checklist_item_id, produccion_id, producto_nombre, tipo_item,
           mee_codigo, descripcion, cantidad_unidades, decoracion_tipo,
           fecha_objetivo, estado, solicitado_por, observaciones)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pendiente', ?, ?)
    """, (item_id, item[1], item[2] or '', item[3] or '',
          item[7] or item[10] or '', item[4] or '', cantidad,
          item[8] or '', fecha_obj, user, observ))
    sol_id = cur.lastrowid

    # Actualizar el item: vincular + estado='solicitado'
    c.execute("""
        UPDATE produccion_checklist SET
          solicitud_produccion_id=?,
          estado='solicitado',
          fecha_solicitud=datetime('now', '-5 hours'),
          actualizado_at=datetime('now', '-5 hours')
        WHERE id=?
    """, (sol_id, item_id))
    conn.commit()
    return jsonify({'ok': True, 'solicitud_id': sol_id, 'item_id': item_id,
                    'mensaje': f'Solicitud {sol_id} enviada a Catalina'})


@bp.route('/api/compras/solicitudes-produccion', methods=['GET'])
def solicitudes_compra_anticipada_list():
    """Lista las solicitudes pendientes (queue de Catalina).

    Querystring: ?estado=pendiente|decidida|todas (default pendiente)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    estado = request.args.get('estado', 'pendiente')
    conn = get_db(); c = conn.cursor()
    where = ""
    params = ()
    if estado != 'todas':
        where = "WHERE sp.estado=?"
        params = (estado,)
    rows = c.execute(f"""
        SELECT sp.id, sp.producto_nombre, sp.tipo_item, sp.mee_codigo,
               sp.descripcion, sp.cantidad_unidades, sp.decoracion_tipo,
               sp.fecha_objetivo, sp.estado, sp.decision, sp.fecha_creacion,
               sp.solicitado_por, sp.observaciones,
               COALESCE(m.stock_actual, 0) as stock_actual,
               COALESCE(m.proveedor, '') as proveedor_default,
               sp.proveedor, sp.oc_numero, sp.tarea_operativa_id
        FROM solicitudes_compra_anticipada sp
        LEFT JOIN maestro_mee m ON m.codigo = sp.mee_codigo
        {where}
        ORDER BY
          CASE sp.estado WHEN 'pendiente' THEN 1 WHEN 'decidida' THEN 2 ELSE 3 END,
          sp.fecha_objetivo ASC, sp.fecha_creacion DESC
        LIMIT 300
    """, params).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    # Sugerencia de ruta para cada solicitud
    for it in items:
        tipo = it.get('tipo_item') or ''
        cant = float(it.get('cantidad_unidades') or 0)
        stock = float(it.get('stock_actual') or 0)
        if tipo in ('etiqueta_frontal', 'etiqueta_posterior', 'etiqueta_lateral', 'instructivo'):
            it['ruta_sugerida'] = 'oc'  # etiquetas siempre OC
        elif it.get('decoracion_tipo') in ('serigrafia', 'tampografia'):
            it['ruta_sugerida'] = 'serigrafia' if it['decoracion_tipo']=='serigrafia' else 'tampografia'
        elif stock >= cant and stock > 0:
            it['ruta_sugerida'] = 'inventario'
        else:
            it['ruta_sugerida'] = 'oc'
    pendientes = sum(1 for x in items if x.get('estado')=='pendiente')
    return jsonify({'items': items, 'total': len(items), 'pendientes': pendientes})


# Alias grupales: cuando una tarea se asigna a un grupo (ej. "operarios"),
# expandir a la lista real de usernames. Los grupos pueden solaparse.
_GRUPOS_USUARIOS = {
    'operarios': ['luz', 'miguel', 'felipe', 'valentina'],  # planta operacion
    'jefes':     ['luz', 'daniela', 'hernando'],
    'compras':   ['catalina'],
    'gerencia':  ['sebastian', 'alejandro'],
    'rrhh':      ['evelin', 'gisseth'],
    'todos':     ['sebastian', 'alejandro', 'catalina', 'luz', 'daniela',
                  'hernando', 'miguel', 'felipe', 'valentina', 'mayra',
                  'evelin', 'gisseth', 'jefferson'],
}


def _resolver_emails_asignados(asignado_a_csv):
    """Convierte un CSV de usernames/grupos en lista de emails unicos.

    Ejemplo:
      'luz,operarios' → ['luz@..', 'miguel@..', 'felipe@..', 'valentina@..']
      (luz aparece en ambos, dedup automatico)

    Si un username no tiene email configurado en USER_EMAILS, simplemente
    se omite (no falla). Esto permite que la app funcione mientras
    Sebastian va llenando los EMAIL_USERS env vars en Render.
    """
    if not asignado_a_csv:
        return []
    try:
        from config import USER_EMAILS as _UE
    except Exception:
        return []
    raw = [t.strip().lower() for t in str(asignado_a_csv).split(',') if t.strip()]
    usernames = []
    for token in raw:
        if token in _GRUPOS_USUARIOS:
            usernames.extend(_GRUPOS_USUARIOS[token])
        else:
            usernames.append(token)
    # dedup preservando orden
    seen = set()
    emails = []
    for u in usernames:
        em = (_UE.get(u, '') or '').strip()
        if em and em.lower() not in seen:
            emails.append(em)
            seen.add(em.lower())
    return emails


def _notificar_tarea_operativa(tarea_id):
    """Envia email a los asignados de una tarea operativa recien creada.

    Lee la tarea (en su propia conexion para no chocar con el commit del
    caller), resuelve emails de los asignados via _resolver_emails_asignados,
    y dispara el envio en background thread (no bloquea la respuesta HTTP
    al usuario que decidio).

    Falla silenciosa: si no hay SMTP configurado, no hay emails resueltos,
    o falla el envio, simplemente loggea y sigue. La tarea ya esta en BD.
    """
    import threading
    def _worker():
        try:
            from config import DB_PATH
            import sqlite3 as _sql
            con = db_connect(timeout=30)
            try:
                row = con.execute("""
                    SELECT id, titulo, descripcion, tipo, asignado_a,
                           fecha_objetivo, cantidad, mee_codigo,
                           producto_relacionado, creado_por
                    FROM tareas_operativas WHERE id=?
                """, (tarea_id,)).fetchone()
            finally:
                con.close()
            if not row:
                return
            (_id, titulo, descripcion, tipo, asignado_a, fecha_obj,
             cantidad, mee_codigo, producto, creado_por) = row
            # Push notif in-app a cada asignado (ANTES del email — feedback inmediato)
            try:
                from blueprints.notif import push_notif_multi
                lista = []
                if asignado_a:
                    raw = [t.strip().lower() for t in str(asignado_a).split(',') if t.strip()]
                    GROUP_ALIASES = {'operarios': ['mayerlin','camilo','milton','sebastian_murillo']}
                    for tok in raw:
                        if tok in GROUP_ALIASES:
                            lista.extend(GROUP_ALIASES[tok])
                        else:
                            lista.append(tok)
                if lista:
                    cuerpo = (descripcion or '')[:140]
                    if fecha_obj:
                        cuerpo += f' · 📅 {fecha_obj}'
                    push_notif_multi(
                        lista, 'tarea_asignada', f'Tarea: {titulo}',
                        body=cuerpo,
                        link='/inventarios#tareas',
                        remitente=(creado_por or 'sistema'),
                        importante=bool(fecha_obj)
                    )
            except Exception as _e:
                logger.warning('push_notif tarea_operativa fallo: %s', _e)
            # ── Email DESACTIVADO · Sebastián 16-may-2026 ──────────────────
            # El email por CADA tarea operativa generaba cientos de correos
            # al día. La notificación in-app (campana · push_notif arriba) ya
            # avisa al asignado. Para reactivar el email: quitar este return.
            return
            emails = _resolver_emails_asignados(asignado_a)
            if not emails:
                return  # nadie con email configurado, no hay nada que enviar
            # Armar email HTML claro y accionable
            urgencia_color = '#dc2626' if fecha_obj else '#1e40af'
            html = f"""
            <div style="font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:560px;margin:0 auto;background:#fff;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden">
              <div style="background:#0f172a;color:#fff;padding:18px 22px">
                <div style="font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:.6px">Nueva tarea operativa · EOS Inventarios</div>
                <h2 style="margin:6px 0 0;font-size:18px">{_html_escape(titulo)}</h2>
              </div>
              <div style="padding:18px 22px;color:#1e293b;font-size:14px;line-height:1.55">
                <p style="margin:0 0 12px">{_html_escape(descripcion)}</p>
                <table style="border-collapse:collapse;font-size:13px;margin-top:14px">
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Cantidad</td><td style="font-weight:700">{int(cantidad or 0):,} und</td></tr>
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Producto</td><td>{_html_escape(producto or '—')}</td></tr>
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Material</td><td style="font-family:monospace">{_html_escape(mee_codigo or '—')}</td></tr>
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Fecha objetivo</td><td style="color:{urgencia_color};font-weight:700">{_html_escape(fecha_obj or 'sin fecha')}</td></tr>
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Tipo</td><td><code style="background:#f1f5f9;padding:2px 6px;border-radius:4px;font-size:11px">{_html_escape(tipo or '')}</code></td></tr>
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Asignada a</td><td>{_html_escape(asignado_a or '')}</td></tr>
                  <tr><td style="padding:4px 12px 4px 0;color:#64748b">Creada por</td><td>{_html_escape(creado_por or '')}</td></tr>
                </table>
                <div style="margin-top:18px;padding:12px 14px;background:#f0fdf4;border-left:3px solid #16a34a;border-radius:6px;color:#166534;font-size:13px">
                  Marca <b>Completar</b> en el modulo cuando termines:<br>
                  <a href="{APP_BASE_URL}/inventarios" style="color:#0f766e;font-weight:600">Planta → Programación → Tareas operativas</a>
                </div>
              </div>
              <div style="background:#f8fafc;padding:12px 22px;font-size:11px;color:#64748b;border-top:1px solid #e2e8f0">
                Esta notificacion se envio automaticamente. No respondas a este correo.
              </div>
            </div>
            """
            asunto = f"[EOS] Tarea pendiente: {titulo}"
            try:
                from notificaciones import SistemaNotificaciones
                import os as _os
                sn = SistemaNotificaciones(
                    email_remitente=_os.environ.get('EMAIL_REMITENTE', ''),
                    contraseña=_os.environ.get('EMAIL_PASSWORD', ''),
                    smtp_server=_os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
                    smtp_port=int(_os.environ.get('SMTP_PORT', '587')),
                )
                sn._enviar_email(asunto, html, destinatarios=emails)
            except Exception:
                import logging
                logging.getLogger('programacion').warning(
                    'notificacion tarea operativa #%s fallo', tarea_id, exc_info=True
                )
        except Exception:
            pass

    threading.Thread(target=_worker, daemon=True).start()


def _html_escape(s):
    """HTML-escape minimo para meter strings en templates de email."""
    if s is None:
        return ''
    return (str(s)
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;'))


@bp.route('/api/compras/solicitudes-produccion/<int:sol_id>/decidir', methods=['POST'])
def solicitudes_compra_anticipada_decidir(sol_id):
    """Catalina decide ruta para una solicitud:
       inventario | oc | serigrafia | tampografia | etiqueta_adhesiva

    Body: {decision, proveedor?, fecha_objetivo?, observaciones?, asignado_a? (CSV)}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    decision = (d.get('decision') or '').strip()
    if decision not in ('inventario', 'oc', 'serigrafia', 'tampografia', 'etiqueta_adhesiva'):
        return jsonify({'error': 'decision invalida'}), 400
    proveedor = (d.get('proveedor') or '').strip()
    fecha_obj = (d.get('fecha_objetivo') or '').strip()
    asignado_a = (d.get('asignado_a') or '').strip()
    obs = (d.get('observaciones') or '').strip()

    conn = get_db(); c = conn.cursor()
    sol = c.execute("""
        SELECT id, checklist_item_id, producto_nombre, tipo_item, mee_codigo,
               descripcion, cantidad_unidades, decoracion_tipo, fecha_objetivo
        FROM solicitudes_compra_anticipada WHERE id=?
    """, (sol_id,)).fetchone()
    if not sol:
        return jsonify({'error': 'Solicitud no encontrada'}), 404

    tarea_id = None
    oc_num = ''

    # Si la decision es 'inventario' Y hay decoracion (serigrafia/tampografia),
    # generar tarea operativa para sacar envases de bodega y mandarlos a decorar
    if decision in ('serigrafia', 'tampografia'):
        # Genera tarea operativa: sacar envases blancos de bodega para enviar
        titulo = f"Sacar {int(sol[6])} envases para {decision} - {sol[2]}"
        descripcion = (
            f"Sacar de bodega {int(sol[6])} unidades de {sol[5] or sol[4]} "
            f"para enviar a {decision} con proveedor {proveedor or 'por definir'}. "
            f"Producto destino: {sol[2]}. {obs}"
        )
        cur = c.execute("""
            INSERT INTO tareas_operativas
              (titulo, descripcion, tipo, producto_relacionado, mee_codigo,
               cantidad, asignado_a, fecha_objetivo, estado,
               origen_tipo, origen_id, creado_por)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pendiente', 'solicitud_produccion', ?, ?)
        """, (titulo, descripcion,
              f'sacar_envases_{decision}',
              sol[2], sol[4] or '', sol[6],
              asignado_a or 'luz,operarios',
              fecha_obj or sol[8] or '',
              sol_id, user))
        tarea_id = cur.lastrowid
    elif decision == 'inventario':
        titulo = f"Sacar {int(sol[6])} de bodega - {sol[5] or sol[4]}"
        descripcion = (
            f"Producción {sol[2]} requiere {int(sol[6])} unidades de "
            f"{sol[5] or sol[4]}. Sacar de bodega y dejar listo. {obs}"
        )
        cur = c.execute("""
            INSERT INTO tareas_operativas
              (titulo, descripcion, tipo, producto_relacionado, mee_codigo,
               cantidad, asignado_a, fecha_objetivo, estado,
               origen_tipo, origen_id, creado_por)
            VALUES (?, ?, 'sacar_inventario', ?, ?, ?, ?, ?, 'pendiente',
                    'solicitud_produccion', ?, ?)
        """, (titulo, descripcion, sol[2], sol[4] or '', sol[6],
              asignado_a or 'luz,operarios',
              fecha_obj or sol[8] or '',
              sol_id, user))
        tarea_id = cur.lastrowid
    elif decision in ('oc', 'etiqueta_adhesiva'):
        # Auto-crear OC en estado Borrador con los datos de la solicitud,
        # linkeada bidireccionalmente con el item del checklist y la
        # solicitud anticipada. Asi Catalina solo entra a /compras a
        # ajustar precios y autorizarla — no tiene que recrear el item.
        # Cuando la OC se reciba en /recepcion, recibir_oc cierra el
        # item del checklist automaticamente via oc_numero.
        try:
            from datetime import datetime as _dt
            year = _dt.now().year
            row = c.execute(
                "SELECT COALESCE(MAX(CAST(SUBSTR(numero_oc,9) AS INTEGER)),0) "
                "FROM ordenes_compra WHERE numero_oc LIKE ?",
                (f'OC-{year}-%',)
            ).fetchone()
            seq = (row[0] or 0) + 1
            oc_num = f'OC-{year}-{seq:04d}'
            categoria = 'MEE'  # envases/etiquetas/decoracion siempre son MEE
            obs_oc = (
                f'Auto-generada desde Solicitud #{sol_id} (decision: {decision}). '
                f'Producto destino: {sol[2]}. '
                f'{obs}'
            ).strip()
            c.execute("""
                INSERT INTO ordenes_compra
                  (numero_oc, fecha, estado, proveedor, observaciones,
                   creado_por, fecha_entrega_est, categoria)
                VALUES (?, ?, 'Borrador', ?, ?, ?, ?, ?)
            """, (oc_num, _dt.now().isoformat(), proveedor or 'POR DEFINIR',
                  obs_oc, user, fecha_obj or sol[8] or '', categoria))
            # Item de la OC con la cantidad del checklist
            c.execute("""
                INSERT INTO ordenes_compra_items
                  (numero_oc, codigo_mp, nombre_mp, cantidad_g,
                   precio_unitario, subtotal)
                VALUES (?, ?, ?, ?, 0, 0)
            """, (oc_num, sol[4] or '', sol[5] or sol[4] or '', sol[6] or 0))
            # Linkear bidireccionalmente
            c.execute(
                "UPDATE solicitudes_compra_anticipada SET oc_numero=? WHERE id=?",
                (oc_num, sol_id)
            )
            c.execute(
                "UPDATE produccion_checklist SET oc_numero=? WHERE id=?",
                (oc_num, sol[1])
            )
            # Crear proveedor si no existe (alimenta catalogo)
            if proveedor:
                exists = c.execute(
                    "SELECT 1 FROM proveedores WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?))",
                    (proveedor,)
                ).fetchone()
                if not exists:
                    try:
                        c.execute("""
                            INSERT INTO proveedores
                              (nombre, categoria, fecha_creacion, activo)
                            VALUES (?, ?, ?, 1)
                        """, (proveedor, categoria, _dt.now().isoformat()))
                        try:
                            from audit_helpers import audit_log as _al
                            _al(c, usuario=session.get('compras_user', 'sistema'),
                                accion='CREAR_PROVEEDOR', tabla='proveedores',
                                registro_id=c.lastrowid,
                                despues={'nombre': proveedor[:200],
                                          'categoria': categoria,
                                          'origen': 'auto_planta_oc'},
                                detalle=f"Auto-creado desde planta al generar OC")
                        except Exception:
                            pass
                    except Exception:
                        pass
            oc_num_creada = oc_num
            oc_num = oc_num_creada  # para el return
        except Exception as _e:
            # Si falla el auto-create, queda como antes (Catalina crea manual)
            oc_num = ''

    # Actualizar la solicitud
    c.execute("""
        UPDATE solicitudes_compra_anticipada SET
          estado='decidida', decision=?, decidido_por=?,
          fecha_decision=datetime('now', '-5 hours'),
          proveedor=?, tarea_operativa_id=?, observaciones=?
        WHERE id=?
    """, (decision, user, proveedor, tarea_id, obs, sol_id))
    # Marcar el item del checklist:
    #   inventario  → en_transito (ya existe, esperando que el operario lo saque)
    #   oc/etiqueta_adhesiva → solicitado (con oc_numero linkeada para cierre auto)
    #   serigrafia/tampografia → solicitado (esperando tarea operativa)
    c.execute("""
        UPDATE produccion_checklist SET
          estado=CASE WHEN ? IN ('inventario') THEN 'en_transito'
                      ELSE 'solicitado' END,
          actualizado_at=datetime('now', '-5 hours')
        WHERE id=?
    """, (decision, sol[1]))
    conn.commit()
    # Notificar a los asignados (background thread, falla silenciosa).
    # Esto cierra la queja "no me esta llegando la alerta" — antes la tarea
    # se creaba en BD pero nadie recibia email, solo aparecia en la pantalla
    # de Planta si alguien entraba a verla.
    if tarea_id:
        try:
            _notificar_tarea_operativa(tarea_id)
        except Exception:
            pass
    msg_extra = ''
    if tarea_id:
        msg_extra = f' (tarea operativa #{tarea_id} creada · email enviado a asignados)'
    elif oc_num:
        msg_extra = f' (OC {oc_num} creada en Borrador — Catalina ajusta precios y autoriza)'
    return jsonify({
        'ok': True, 'solicitud_id': sol_id, 'decision': decision,
        'tarea_operativa_id': tarea_id,
        'oc_numero': oc_num,
        'mensaje': f'Solicitud {sol_id} marcada como {decision}' + msg_extra,
    })


@bp.route('/api/tareas-operativas', methods=['GET'])
def tareas_operativas_list():
    """Lista de tareas operativas. Querystring:
       ?usuario=<username> (filtra por asignado_a)
       ?estado=pendiente|completada|todas (default pendiente+en_progreso)
       ?contexto=planta|mias|todas (Sebastian 29-abr-2026: el caso
        "Cargar influencers" salia en /planta porque era una tarea
        chat_asignacion para Jeferson — la asignacion era grupal).
        - planta: solo tareas fisicas de planta (excluye chat_asignacion)
        - mias: solo asignadas al usuario actual
        - todas (default): backward-compatible
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    me = (session.get('compras_user', '') or '').strip().lower()
    usuario = (request.args.get('usuario') or '').strip().lower()
    estado_q = (request.args.get('estado') or '').strip()
    contexto = (request.args.get('contexto') or '').strip().lower()
    conn = get_db(); c = conn.cursor()
    where = []
    params = []
    if estado_q == 'todas':
        pass
    elif estado_q in ('pendiente', 'en_progreso', 'completada', 'cancelada'):
        where.append("estado=?"); params.append(estado_q)
    else:
        where.append("estado IN ('pendiente','en_progreso')")
    if usuario:
        where.append("(LOWER(asignado_a) LIKE ? OR asignado_a='')")
        params.append(f"%{usuario}%")
    if contexto == 'planta':
        # Solo tipos fisicos de planta. chat_asignacion es generica entre
        # cualquier par de usuarios y NO pertenece al flujo de planta.
        where.append("(tipo IN ('sacar_envases_serigrafia','sacar_envases_tampografia','sacar_inventario','envasado','etiquetado','recoleccion_mee','recoleccion_mp','dispensacion','limpieza_sala','cambio_lote','muestreo','revision_visual') "
                     "OR origen_tipo IN ('solicitud_produccion','sol_anticipada'))")
        where.append("COALESCE(origen_tipo,'') != 'chat'")
    elif contexto == 'mias':
        # Solo las asignadas a mi (csv split o match exacto). origen_tipo
        # 'chat' tambien aplica — son tareas que me asignaron en chat.
        where.append("(',' || LOWER(REPLACE(COALESCE(asignado_a,''),' ','')) || ',') LIKE ?")
        params.append(f"%,{me},%")
    sql = "SELECT * FROM tareas_operativas"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY fecha_objetivo ASC, fecha_creacion DESC LIMIT 300"
    rows = c.execute(sql, params).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'tareas': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/tareas-operativas/<int:tarea_id>/completar', methods=['POST'])
def tareas_operativas_completar(tarea_id):
    """Operario marca la tarea como completada.

    Cierra la cadena causal completa: tarea operativa → solicitud anticipada
    → item del checklist. Asi cuando el operario marca "ya saque los envases",
    el item del checklist Pre-Produccion pasa a 'recibido' automaticamente y
    el % listo de la card del producto sube sin intervencion manual.

    Body opcional: {observaciones, cantidad_real (si difiere)}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    obs = (d.get('observaciones') or '').strip()
    conn = get_db(); c = conn.cursor()
    cur = c.execute("""
        UPDATE tareas_operativas SET
          estado='completada',
          completado_por=?,
          fecha_completado=datetime('now', '-5 hours'),
          observaciones_cierre=?
        WHERE id=? AND estado IN ('pendiente','en_progreso')
    """, (user, obs, tarea_id))
    if cur.rowcount == 0:
        return jsonify({'error': 'Tarea no encontrada o ya completada'}), 404
    # Marcar la solicitud_produccion origen como completada (si aplica)
    # y propagar al item del checklist como 'recibido'.
    sol_row = c.execute("""
        SELECT id, checklist_item_id FROM solicitudes_compra_anticipada
        WHERE tarea_operativa_id=?
    """, (tarea_id,)).fetchone()
    checklist_item_id = None
    if sol_row:
        sol_id, checklist_item_id = sol_row[0], sol_row[1]
        c.execute(
            "UPDATE solicitudes_compra_anticipada SET estado='completada' WHERE id=?",
            (sol_id,)
        )
        if checklist_item_id:
            c.execute("""
                UPDATE produccion_checklist SET
                  estado='recibido',
                  fecha_recibido=date('now', '-5 hours'),
                  actualizado_at=datetime('now', '-5 hours')
                WHERE id=?
            """, (checklist_item_id,))
    conn.commit()
    return jsonify({
        'ok': True,
        'tarea_id': tarea_id,
        'checklist_item_id': checklist_item_id,
        'mensaje': 'Tarea completada' + (
            ' · checklist actualizado' if checklist_item_id else ''
        ),
    })


@bp.route('/api/tareas-operativas', methods=['POST'])
def tareas_operativas_crear():
    """Crear una tarea operativa manualmente (para jefes / Catalina).

    Body: {titulo, descripcion, tipo, asignado_a, fecha_objetivo,
           producto_relacionado?, cantidad?, mee_codigo?}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    titulo = (d.get('titulo') or '').strip()
    if not titulo:
        return jsonify({'error': 'titulo requerido'}), 400
    conn = get_db(); c = conn.cursor()
    cur = c.execute("""
        INSERT INTO tareas_operativas
          (titulo, descripcion, tipo, producto_relacionado, mee_codigo,
           cantidad, asignado_a, fecha_objetivo, estado,
           origen_tipo, creado_por)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pendiente', 'manual', ?)
    """, (
        titulo, (d.get('descripcion') or '').strip(),
        (d.get('tipo') or 'general').strip(),
        (d.get('producto_relacionado') or '').strip(),
        (d.get('mee_codigo') or '').strip(),
        float(d.get('cantidad') or 0),
        (d.get('asignado_a') or '').strip(),
        (d.get('fecha_objetivo') or '').strip(),
        user
    ))
    conn.commit()
    tarea_id = cur.lastrowid
    # Notificar a los asignados (background, falla silenciosa)
    try:
        _notificar_tarea_operativa(tarea_id)
    except Exception:
        pass
    return jsonify({'ok': True, 'tarea_id': tarea_id})


# ════════════════════════════════════════════════════════════════════════
# PLANTA INTELIGENTE — FASE 0: Presentaciones por SKU
# ════════════════════════════════════════════════════════════════════════
# Sebastian + Alejandro (30-abr-2026): los productos tienen multiples
# presentaciones (suero 30/15/10mL, contornos 15/10mL, maxlash 4.5mL,
# blush balm 6g). Sin esto, "produzcamos suero para 2 meses" es ambiguo.
# Esta tabla se llena via UI o bulk import. NO toca formula_headers.

# Catalogo de categorias estandar con sus presentaciones default segun el
# brief de Alejandro (30-abr-2026). Sirve como sugerencia al crear.
PRESENTACIONES_DEFAULT_POR_CATEGORIA = {
    'limpiador': [
        {'codigo': 'lmp_150ml', 'etiqueta': 'Limpiador 150 mL', 'volumen_ml': 150,
         'envase_codigo': '', 'notas': 'Plástico tapa rosca blanco'}
    ],
    'hidratante': [
        {'codigo': 'hid_50ml', 'etiqueta': 'Hidratante 50 mL airless', 'volumen_ml': 50,
         'envase_codigo': '', 'notas': 'Plástico airless'}
    ],
    'suero': [
        {'codigo': 'sue_30ml', 'etiqueta': 'Suero 30 mL', 'volumen_ml': 30,
         'envase_codigo': '', 'notas': ''},
        {'codigo': 'sue_15ml', 'etiqueta': 'Suero 15 mL', 'volumen_ml': 15,
         'envase_codigo': '', 'notas': ''},
        {'codigo': 'sue_10ml', 'etiqueta': 'Suero 10 mL', 'volumen_ml': 10,
         'envase_codigo': '', 'notas': ''},
    ],
    'contorno_ojos': [
        {'codigo': 'co_15ml', 'etiqueta': 'Contorno 15 mL (multipéptidos/retinal)', 'volumen_ml': 15,
         'envase_codigo': '', 'notas': 'Multipéptidos y retinal'},
        {'codigo': 'co_10ml', 'etiqueta': 'Contorno 10 mL (cafeína)', 'volumen_ml': 10,
         'envase_codigo': '', 'notas': 'Cafeína'},
    ],
    'maxlash': [
        {'codigo': 'mx_45ml', 'etiqueta': 'Maxlash 4.5 mL', 'volumen_ml': 4.5,
         'envase_codigo': '', 'notas': 'Suero cejas y pestañas'}
    ],
    'blush_balm': [
        {'codigo': 'bb_6g', 'etiqueta': 'Blush Balm 6 g', 'volumen_ml': None,
         'envase_codigo': '', 'notas': 'Peso 6g'}
    ],
}


@bp.route('/api/planta/presentaciones', methods=['GET'])
def planta_presentaciones_list():
    """Lista presentaciones registradas. Querystring opcional:
       ?producto=<nombre> filtra por producto exacto
       ?categoria=<cat> filtra por categoria
       ?activos=1 (default) solo activos
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    where = []
    params = []
    if request.args.get('producto'):
        where.append('producto_nombre = ?')
        params.append(request.args.get('producto').strip())
    if request.args.get('categoria'):
        where.append('categoria = ?')
        params.append(request.args.get('categoria').strip())
    if request.args.get('activos', '1') == '1':
        where.append('activo = 1')
    sql = "SELECT * FROM producto_presentaciones"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY producto_nombre, COALESCE(volumen_ml, 999) DESC, etiqueta LIMIT 2000"
    rows = c.execute(sql, params).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    # Resumen por producto
    productos = {}
    for it in items:
        pn = it['producto_nombre'] or '(sin asignar)'
        productos.setdefault(pn, []).append(it)
    return jsonify({
        'presentaciones': items,
        'total': len(items),
        'por_producto': productos,
        'plantillas_default': PRESENTACIONES_DEFAULT_POR_CATEGORIA,
    })


@bp.route('/api/planta/presentaciones', methods=['POST'])
def planta_presentaciones_crear():
    """Crear presentacion individual.
    Body: {producto_nombre*, categoria, presentacion_codigo*, etiqueta*,
           volumen_ml, peso_g, envase_codigo, factor_g_por_unidad,
           sku_shopify, es_default, notas}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    producto = (d.get('producto_nombre') or '').strip()
    pcode = (d.get('presentacion_codigo') or '').strip()
    etiqueta = (d.get('etiqueta') or '').strip()
    if not producto:
        return jsonify({'error': 'producto_nombre requerido'}), 400
    if not pcode:
        return jsonify({'error': 'presentacion_codigo requerido'}), 400
    if not etiqueta:
        return jsonify({'error': 'etiqueta requerida'}), 400
    conn = get_db(); c = conn.cursor()
    try:
        cur = c.execute("""
            INSERT INTO producto_presentaciones
              (producto_nombre, categoria, presentacion_codigo, etiqueta,
               volumen_ml, peso_g, envase_codigo, factor_g_por_unidad,
               sku_shopify, es_default, activo, notas, actualizado_en)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, datetime('now', '-5 hours'))
        """, (
            producto,
            (d.get('categoria') or '').strip() or None,
            pcode, etiqueta,
            d.get('volumen_ml'),
            d.get('peso_g'),
            (d.get('envase_codigo') or '').strip() or None,
            d.get('factor_g_por_unidad'),
            (d.get('sku_shopify') or '').strip() or None,
            1 if d.get('es_default') else 0,
            (d.get('notas') or '').strip() or None,
        ))
        conn.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})
    except sqlite3.IntegrityError as e:
        return jsonify({'error': f'Presentación ya existe para este producto: {e}'}), 409


@bp.route('/api/planta/presentaciones/<int:pid>', methods=['PUT', 'DELETE'])
def planta_presentaciones_detail(pid):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'DELETE':
        # Soft delete (activo=0). No borramos para preservar referencias historicas.
        c.execute("UPDATE producto_presentaciones SET activo=0, actualizado_en=datetime('now', '-5 hours') WHERE id=?", (pid,))
        conn.commit()
        return jsonify({'ok': True})
    # PUT
    d = request.json or {}
    campos = []
    params = []
    for col in ('etiqueta', 'categoria', 'volumen_ml', 'peso_g', 'envase_codigo',
                'factor_g_por_unidad', 'sku_shopify', 'es_default', 'notas'):
        if col in d:
            campos.append(f'{col} = ?')
            v = d[col]
            if col in ('volumen_ml', 'peso_g', 'factor_g_por_unidad'):
                v = float(v) if v not in (None, '') else None
            elif col == 'es_default':
                v = 1 if v else 0
            elif isinstance(v, str):
                v = v.strip() or None
            params.append(v)
    if not campos:
        return jsonify({'error': 'Sin cambios'}), 400
    campos.append("actualizado_en = datetime('now', '-5 hours')")
    params.append(pid)
    c.execute(f"UPDATE producto_presentaciones SET {', '.join(campos)} WHERE id=?", params)
    conn.commit()
    return jsonify({'ok': True, 'updated': c.rowcount})


@bp.route('/api/planta/presentaciones/bulk-categoria', methods=['POST'])
def planta_presentaciones_bulk_categoria():
    """Aplica las plantillas default de una categoria a un producto especifico.
    Util para arrancar rapido — Alejandro elige producto + categoria y se
    crean las N presentaciones default.
    Body: {producto_nombre*, categoria*}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    producto = (d.get('producto_nombre') or '').strip()
    categoria = (d.get('categoria') or '').strip().lower()
    if not producto or not categoria:
        return jsonify({'error': 'producto_nombre y categoria requeridos'}), 400
    plantillas = PRESENTACIONES_DEFAULT_POR_CATEGORIA.get(categoria)
    if not plantillas:
        return jsonify({
            'error': f'Categoria "{categoria}" no reconocida',
            'categorias_validas': list(PRESENTACIONES_DEFAULT_POR_CATEGORIA.keys())
        }), 400
    conn = get_db(); c = conn.cursor()
    creadas = []
    for p in plantillas:
        try:
            cur = c.execute("""
                INSERT INTO producto_presentaciones
                  (producto_nombre, categoria, presentacion_codigo, etiqueta,
                   volumen_ml, envase_codigo, es_default, activo, notas, actualizado_en)
                VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, datetime('now', '-5 hours'))
            """, (
                producto, categoria, p['codigo'], p['etiqueta'],
                p['volumen_ml'], p.get('envase_codigo') or None,
                1 if len(plantillas) == 1 else 0,
                p.get('notas') or None,
            ))
            creadas.append({'id': cur.lastrowid, 'codigo': p['codigo']})
        except sqlite3.IntegrityError:
            # Ya existia — saltar silenciosamente
            pass
    conn.commit()
    return jsonify({'ok': True, 'creadas': creadas, 'total': len(creadas)})


@bp.route('/api/planta/presentaciones/productos-disponibles', methods=['GET'])
def planta_presentaciones_productos_disponibles():
    """Lista productos de formula_headers + cuántas presentaciones tienen ya."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT fh.producto_nombre,
               fh.lote_size_kg,
               (SELECT COUNT(*) FROM producto_presentaciones pp
                WHERE pp.producto_nombre = fh.producto_nombre AND pp.activo=1) AS n_presentaciones
        FROM formula_headers fh
        ORDER BY fh.producto_nombre
    """).fetchall()
    items = []
    for r in rows:
        items.append({
            'producto_nombre': r[0],
            'lote_size_kg': r[1],
            'n_presentaciones': r[2] or 0,
        })
    return jsonify({'productos': items, 'total': len(items)})


# ════════════════════════════════════════════════════════════════════════
# PLANTA INTELIGENTE — FASE 1: Catálogo de equipos + sugerencia de área
# ════════════════════════════════════════════════════════════════════════
# Sebastian + Alejandro (30-abr-2026): el Excel "LISTADO MAESTRO DE EQUIPOS
# 2026" tiene 104 equipos en 9 areas reales. La migracion 63 los importa.
# Aqui exponemos endpoints para listar, ver y sugerir area al programar
# producciones (cruce capacidad-de-tanque vs tamaño-de-lote).

@bp.route('/api/planta/equipos', methods=['GET'])
def planta_equipos_list():
    """Lista todos los equipos. Querystring opcional:
       ?area=<codigo>  filtra por area
       ?tipo=<tipo>    filtra por tipo (tanque, envasadora, etc.)
       ?con_capacidad=1 solo equipos con capacidad_litros NOT NULL
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    where = ['activo=1']
    params = []
    if request.args.get('area'):
        where.append('area_codigo = ?')
        params.append(request.args.get('area').strip())
    if request.args.get('tipo'):
        where.append('tipo = ?')
        params.append(request.args.get('tipo').strip())
    if request.args.get('con_capacidad') == '1':
        where.append('capacidad_litros IS NOT NULL')
    sql = "SELECT * FROM equipos_planta WHERE " + " AND ".join(where) + " ORDER BY area_codigo, COALESCE(capacidad_litros, 0) DESC, nombre"
    rows = c.execute(sql, params).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    # Resumen por área + por tipo
    por_area = {}
    por_tipo = {}
    for it in items:
        a = it.get('area_codigo') or '—'
        t = it.get('tipo') or 'otro'
        por_area.setdefault(a, []).append(it)
        por_tipo[t] = por_tipo.get(t, 0) + 1
    return jsonify({
        'equipos': items,
        'total': len(items),
        'por_area': por_area,
        'por_tipo': por_tipo,
    })


@bp.route('/api/planta/equipos/<int:eq_id>', methods=['GET', 'PUT', 'DELETE'])
def planta_equipos_detail(eq_id):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'GET':
        r = c.execute("SELECT * FROM equipos_planta WHERE id=?", (eq_id,)).fetchone()
        if not r:
            return jsonify({'error': 'No existe'}), 404
        cols = [d[0] for d in c.description]
        return jsonify({'equipo': dict(zip(cols, r))})
    if request.method == 'DELETE':
        c.execute("UPDATE equipos_planta SET activo=0, actualizado_en=datetime('now', '-5 hours') WHERE id=?", (eq_id,))
        conn.commit()
        return jsonify({'ok': True})
    # PUT
    d = request.json or {}
    campos = []
    params = []
    for col in ('nombre', 'area_codigo', 'tipo', 'capacidad_raw',
                'capacidad_litros', 'capacidad_kg', 'estado_operacional',
                'notas'):
        if col in d:
            campos.append(f'{col} = ?')
            v = d[col]
            if col in ('capacidad_litros', 'capacidad_kg'):
                v = float(v) if v not in (None, '') else None
            elif isinstance(v, str):
                v = v.strip() or None
            params.append(v)
    if not campos:
        return jsonify({'error': 'Sin cambios'}), 400
    campos.append("actualizado_en = datetime('now', '-5 hours')")
    params.append(eq_id)
    c.execute(f"UPDATE equipos_planta SET {', '.join(campos)} WHERE id=?", params)
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/planta/areas/v2', methods=['GET'])
def planta_areas_v2():
    """Listado de areas con equipos asignados. Mas rico que /api/planta/areas."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT a.codigo, a.nombre, a.tipo, a.estado, a.activo,
               a.requiere_limpieza_profunda,
               a.ultima_limpieza_profunda,
               a.puede_producir, a.puede_envasar,
               (SELECT COUNT(*) FROM equipos_planta WHERE area_codigo=a.codigo AND activo=1) AS n_equipos,
               (SELECT MAX(capacidad_litros) FROM equipos_planta
                  WHERE area_codigo=a.codigo AND activo=1 AND tipo='tanque') AS max_tanque_l,
               (SELECT MIN(capacidad_litros) FROM equipos_planta
                  WHERE area_codigo=a.codigo AND activo=1 AND tipo='tanque'
                    AND capacidad_litros IS NOT NULL) AS min_tanque_l
        FROM areas_planta a
        WHERE a.activo=1
        ORDER BY a.orden, a.codigo
    """).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'areas': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/planta/sugerir-area', methods=['POST'])
def planta_sugerir_area():
    """Sugerencia inteligente de área para una producción.

    Body: {producto_nombre*, lote_kg*, presentacion_codigo?}

    Lógica (Sebastian + Alejandro, 30-abr-2026):
      1. Filtrar areas con puede_producir=1 que tengan al menos 1 tanque
         con capacidad >= lote_kg * 1.2 (margen 20%, asumiendo densidad ~1).
      2. De las que pasan, preferir el tanque MAS PEQUEÑO que aguante
         (eficiencia: no usar tanque 400L para lote de 30L).
      3. Score adicional:
         + estado='libre' (sin produccion en curso)
         + cercania al envasado (FAB1->ENV1, FAB2/3->ENV2)
         + ultima_limpieza_profunda reciente (<7 dias)
      4. Devolver lista ordenada con score y razon.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    producto = (d.get('producto_nombre') or '').strip()
    lote_kg = float(d.get('lote_kg') or 0)
    if not producto or lote_kg <= 0:
        return jsonify({'error': 'producto_nombre y lote_kg requeridos'}), 400
    conn = get_db(); c = conn.cursor()

    # Margen de 20% sobre el lote (densidad cercana a agua = 1 g/mL = 1 kg/L)
    capacidad_min_litros = lote_kg * 1.2

    # Buscar tanques candidatos (con capacidad suficiente)
    tanques = c.execute("""
        SELECT id, codigo, nombre, area_codigo, capacidad_litros
        FROM equipos_planta
        WHERE activo=1 AND tipo IN ('tanque','marmita','olla')
          AND capacidad_litros IS NOT NULL
          AND capacidad_litros >= ?
        ORDER BY capacidad_litros ASC
    """, (capacidad_min_litros,)).fetchall()

    # Agrupar por area y tomar el mas pequeño que aguante
    por_area = {}
    for t in tanques:
        a = t[3]
        if a not in por_area or t[4] < por_area[a]['capacidad_litros']:
            por_area[a] = {
                'tanque_id': t[0],
                'tanque_codigo': t[1],
                'tanque_nombre': t[2],
                'capacidad_litros': t[4],
            }

    # Cargar info de areas candidatas
    sugerencias = []
    for area_codigo, tanque_info in por_area.items():
        area_row = c.execute("""
            SELECT codigo, nombre, estado, puede_producir, puede_envasar,
                   requiere_limpieza_profunda, ultima_limpieza_profunda
            FROM areas_planta WHERE codigo=? AND activo=1
        """, (area_codigo,)).fetchone()
        if not area_row:
            continue
        if not area_row[3]:  # puede_producir
            continue
        score = 0
        razones = []
        # Eficiencia: tanque pequeño que aguanta
        utilizacion = (lote_kg / tanque_info['capacidad_litros']) * 100
        score += 50 if utilizacion >= 50 else 30
        razones.append(f"Tanque {tanque_info['capacidad_litros']:.0f}L · uso {utilizacion:.0f}%")
        # Estado libre
        if area_row[2] == 'libre':
            score += 20
            razones.append('Sala libre')
        elif area_row[2] == 'ocupada':
            score -= 10
            razones.append('Sala ocupada')
        # Cercania a envasado (regla suave Fab1->Env1, Fab2/3->Env2)
        env_cercano = ''
        if area_codigo in ('FAB1', 'PROD1'):
            env_cercano = 'ENV1'
        elif area_codigo in ('FAB2', 'PROD2', 'FAB3', 'PROD3'):
            env_cercano = 'ENV2'
        # Limpieza reciente
        if area_row[5]:  # requiere limpieza profunda
            if area_row[6]:
                from datetime import datetime as _dt
                try:
                    last = _dt.fromisoformat((area_row[6] or '').split('.')[0])
                    dias = (_dt.now() - last).days
                    if dias <= 7:
                        score += 10
                        razones.append(f"Limpieza profunda hace {dias}d")
                    elif dias > 14:
                        score -= 15
                        razones.append(f"⚠ Limpieza profunda hace {dias}d (>14d)")
                except Exception:
                    pass
            else:
                score -= 5
                razones.append('Sin registro de limpieza profunda')
        sugerencias.append({
            'area_codigo': area_codigo,
            'area_nombre': area_row[1],
            'tanque': tanque_info,
            'estado_actual': area_row[2],
            'envasado_sugerido': env_cercano,
            'utilizacion_pct': round(utilizacion, 1),
            'score': score,
            'razones': razones,
        })

    sugerencias.sort(key=lambda x: x['score'], reverse=True)

    return jsonify({
        'producto_nombre': producto,
        'lote_kg': lote_kg,
        'capacidad_minima_litros': capacidad_min_litros,
        'sugerencias': sugerencias,
        'recomendada': sugerencias[0] if sugerencias else None,
        'mensaje': (
            f'No hay área con tanque ≥ {capacidad_min_litros:.0f}L disponible. '
            f'Considera dividir el lote o revisar el catálogo de equipos.'
            if not sugerencias else
            f'{len(sugerencias)} área(s) candidata(s). '
            f'Recomendada: {sugerencias[0]["area_nombre"]} '
            f'({sugerencias[0]["tanque"]["capacidad_litros"]:.0f}L).'
        ),
    })


# ════════════════════════════════════════════════════════════════════════
# AUTO-ASIGNACIÓN IA · área + envasado + operarios + limpieza mismo día
# ════════════════════════════════════════════════════════════════════════
# Sebastián (1-may-2026): "TODOS rotan, queden limpias el mismo día, IA
# que sepa qué usar — si producción 200kg debe decir producción donde está
# marmita 250 litros, si lo haces automático sería maravilloso".


def _siguiente_operario_rotando(c, rol, candidatos_ids):
    """Round-robin estricto entre los operarios candidatos.

    Lee rotacion_operarios_state[rol], devuelve el próximo en la lista.
    Si el último estaba al final → vuelve al primero.

    BUG-15 fix · 19-may-2026 audit Planta PERFECTA: si el último ya no
    está en candidatos (operario inactivado), antes devolvíamos
    `candidatos_ids[0]` (el id menor) → la rotación se sesgaba hacia el
    id mínimo cada vez que se inactivaba alguien. Ahora buscamos el id
    inmediatamente superior a `ultimo`; si todos son menores, recién
    volvemos al primero.
    """
    if not candidatos_ids:
        return None
    try:
        row = c.execute(
            "SELECT ultimo_operario_id FROM rotacion_operarios_state WHERE rol=?",
            (rol,)
        ).fetchone()
        ultimo = row[0] if row else None
    except Exception:
        ultimo = None
    if not ultimo:
        return candidatos_ids[0]
    if ultimo in candidatos_ids:
        try:
            idx = candidatos_ids.index(ultimo)
            return candidatos_ids[(idx + 1) % len(candidatos_ids)]
        except ValueError:
            return candidatos_ids[0]
    # ultimo no está en candidatos · buscar el siguiente id mayor a ultimo;
    # si no hay (todos los candidatos < ultimo), wrap a candidatos_ids[0]
    siguiente_mayor = next((c_id for c_id in candidatos_ids if c_id > ultimo),
                            None)
    return siguiente_mayor if siguiente_mayor is not None else candidatos_ids[0]


def _registrar_rotacion(c, rol, operario_id, user='auto-ia'):
    """Persiste la asignación para que la próxima iteración rote al siguiente."""
    try:
        c.execute("""
            INSERT INTO rotacion_operarios_state (rol, ultimo_operario_id, ultimo_asignado_at, actualizado_por)
            VALUES (?, ?, datetime('now', '-5 hours'), ?)
            ON CONFLICT(rol) DO UPDATE SET
              ultimo_operario_id=excluded.ultimo_operario_id,
              ultimo_asignado_at=datetime('now', '-5 hours'),
              actualizado_por=excluded.actualizado_por
        """, (rol, operario_id, user))
    except Exception:
        # Fallback si SQLite vieja sin ON CONFLICT
        c.execute(
            "UPDATE rotacion_operarios_state SET ultimo_operario_id=?, ultimo_asignado_at=datetime('now', '-5 hours'), actualizado_por=? WHERE rol=?",
            (operario_id, user, rol)
        )


def _operarios_libres_en_dia(c, fecha_iso):
    """Operarios que NO están ya asignados a otra producción ese día.
    Sebastián 1-may-2026: excluye jefes de producción (Luis Enrique no rota)."""
    rows = c.execute("""
        SELECT id, nombre, COALESCE(apellido,''), rol_predeterminado
        FROM operarios_planta
        WHERE COALESCE(activo,1)=1
          AND COALESCE(es_jefe_produccion,0)=0
        ORDER BY id
    """).fetchall()
    todos = [(r[0], (r[1]+' '+r[2]).strip(), r[3] or '') for r in rows]
    # Excluir operarios ya asignados ese día
    ocupados_rows = c.execute("""
        SELECT operario_dispensacion_id, operario_elaboracion_id,
               operario_envasado_id, operario_acondicionamiento_id
        FROM produccion_programada
        WHERE date(fecha_programada) = ?
          AND COALESCE(estado, 'programado') != 'cancelado'
    """, (fecha_iso,)).fetchall()
    ocupados = set()
    for r in ocupados_rows:
        for x in r:
            if x: ocupados.add(x)
    return [(oid, nom, rol) for (oid, nom, rol) in todos if oid not in ocupados], todos


def _auto_asignar_operarios(c, produccion_id, fecha_iso, user='auto-ia'):
    """Asigna 4 roles ponderando preferencias + AVOID double-booking del día.
    Sebastián 1-may-2026 (round 2): 'mismo operario en 4 producciones distintas
    el mismo día' es físicamente imposible · ahora la IA evita asignar operarios
    que YA están en otras producciones del mismo día (por rol).

    Lógica:
    1. Pool = 4 operarios activos no-jefe
    2. Carga global: operarios_ya_usados_por_rol[rol] = set(op_ids) de OTRAS
       producciones del mismo día
    3. Para cada rol, intenta primero: pool - usados_en_esta_prod - globalmente_usados_para_rol
    4. Si vacío, fallback: pool - usados_en_esta_prod (acepta double-book)
    5. Aplica afinidad ponderada para escoger el final
    """
    import hashlib
    # Pool = operarios activos NO jefes. Se trae también fija_en_dispensacion
    # para enforzar la regla dura del CEO (Mayerlin SOLO dispensa).
    todos_rows = c.execute("""
        SELECT id, COALESCE(rol_predeterminado, ''),
               COALESCE(fija_en_dispensacion, 0)
        FROM operarios_planta
        WHERE COALESCE(activo, 1) = 1
          AND COALESCE(es_jefe_produccion, 0) = 0
        ORDER BY id
    """).fetchall()
    pool = [(r[0], (r[1] or '').strip().lower(), bool(r[2])) for r in todos_rows]
    if not pool:
        return None

    # Producto+fecha para hash determinístico
    prod_row = c.execute(
        "SELECT producto FROM produccion_programada WHERE id=?",
        (produccion_id,)
    ).fetchone()
    producto = (prod_row[0] or '') if prod_row else ''

    # ── Global day load: operarios ya asignados en OTRAS producciones del día
    # Sebastián 1-may-2026: evita 'Camilo en 4 producciones · Lunes'
    globalmente_usados = {
        'dispensacion': set(),
        'elaboracion': set(),
        'envasado': set(),
        'acondicionamiento': set(),
    }
    try:
        otras_rows = c.execute("""
            SELECT operario_dispensacion_id, operario_elaboracion_id,
                   operario_envasado_id, operario_acondicionamiento_id
            FROM produccion_programada
            WHERE date(fecha_programada) = ?
              AND id != ?
              AND COALESCE(estado, 'programado') NOT IN ('cancelado','completado')
        """, (fecha_iso, produccion_id)).fetchall()
        for r in otras_rows:
            if r[0]: globalmente_usados['dispensacion'].add(r[0])
            if r[1]: globalmente_usados['elaboracion'].add(r[1])
            if r[2]: globalmente_usados['envasado'].add(r[2])
            if r[3]: globalmente_usados['acondicionamiento'].add(r[3])
    except Exception as _e:
        log = logging.getLogger('inventario.programacion')
        log.warning('global day load fallo prod=%s fecha=%s: %s',
                    produccion_id, fecha_iso, _e)

    # Sebastián 1-may-2026 audit: leer AFINIDAD de tabla rol_afinidad_config
    # (migración 81). Antes hardcoded duplicado entre auto_plan.py y este archivo.
    try:
        from blueprints.auto_plan import _cargar_afinidad
        AFINIDAD = _cargar_afinidad(c)
    except Exception as _e:
        log = logging.getLogger('inventario.programacion')
        log.warning('_cargar_afinidad fallback hardcoded: %s', _e)
        AFINIDAD = {
            'dispensacion': {'dispensacion': 4, 'elaboracion': 1, 'envasado': 1,
                              'acondicionamiento': 1, 'todero': 2},
            'elaboracion':  {'dispensacion': 1, 'elaboracion': 4, 'envasado': 1,
                              'acondicionamiento': 1, 'todero': 2},
            'envasado':     {'envasado': 4, 'dispensacion': 1, 'elaboracion': 1,
                              'acondicionamiento': 1, 'todero': 2},
            'acondicionamiento': {'acondicionamiento': 4, 'envasado': 2, 'dispensacion': 1,
                                    'elaboracion': 1, 'todero': 2},
        }

    def _hash_rot(rol):
        s = f'{producto}|{fecha_iso}|{rol}'
        return int(hashlib.md5(s.encode('utf-8')).hexdigest()[:8], 16)

    # Pre-segmentación: fijos vs móviles
    pool_fijos = [(oid, rp) for (oid, rp, fija) in pool if fija]
    pool_moviles = [(oid, rp) for (oid, rp, fija) in pool if not fija]

    asignaciones = {}
    usados = set()
    for rol in ('dispensacion', 'elaboracion', 'envasado', 'acondicionamiento'):
        # Regla dura: roles ≠ dispensación NUNCA reciben operarios fija_en_dispensacion
        if rol == 'dispensacion':
            base_pool = pool_moviles + pool_fijos  # móviles primero, fijos como respaldo
            preferido_fijo = [c for c in pool_fijos if c[0] not in usados]
        else:
            base_pool = pool_moviles
            preferido_fijo = []

        # Para dispensación con fijo disponible: forzar fijo (rotación entre fijos por hash)
        if rol == 'dispensacion' and preferido_fijo:
            idx = _hash_rot(rol) % len(preferido_fijo)
            elegido = preferido_fijo[idx][0]
            asignaciones[rol] = elegido
            usados.add(elegido)
            globalmente_usados.get(rol, set()).add(elegido)
            _registrar_rotacion(c, rol, elegido, user)
            continue

        # 1ra preferencia: NO en esta producción Y NO en otras producciones (mismo rol)
        candidatos_strict = [
            (oid, rp) for (oid, rp) in base_pool
            if oid not in usados and oid not in globalmente_usados.get(rol, set())
        ]
        # Fallback 1: solo NO en esta producción
        candidatos_fallback = [(oid, rp) for (oid, rp) in base_pool if oid not in usados]
        # Fallback 2 (último recurso): base_pool completo
        candidatos = candidatos_strict or candidatos_fallback or base_pool
        if not candidatos:
            # Sin base_pool (caso pool=todos fijos y rol≠disp) → registrar warning,
            # NO romper la regla dura. Producción queda con NULL en este rol.
            log = logging.getLogger('inventario.programacion')
            log.warning('rol %s sin candidatos en prod=%s fecha=%s (todos fijos?)',
                        rol, produccion_id, fecha_iso)
            continue
        afin = AFINIDAD.get(rol, {})
        weighted = [(afin.get(rp, 1) if rp else 1, oid) for (oid, rp) in candidatos]
        total = sum(w for w, _ in weighted)
        if total <= 0:
            elegido = candidatos[_hash_rot(rol) % len(candidatos)][0]
        else:
            target = _hash_rot(rol) % total
            cumulative = 0
            elegido = weighted[0][1]
            for w, oid in weighted:
                cumulative += w
                if target < cumulative:
                    elegido = oid
                    break
        asignaciones[rol] = elegido
        usados.add(elegido)
        globalmente_usados.get(rol, set()).add(elegido)
        _registrar_rotacion(c, rol, elegido, user)

    # BUG-11 fix · 19-may-2026: validar que TODOS los 4 roles tengan operario
    # asignado ANTES de tocar la BD. Si pool_moviles está vacío (caso extremo:
    # todos los activos son fijos en dispensación o jefes), algún rol queda
    # sin candidato → producción con roles NULL parcial = estado inconsistente.
    # Antes: el UPDATE con COALESCE dejaba NULL los faltantes sin avisar.
    # Ahora: abortamos sin tocar la BD si falta cualquier rol.
    roles_esperados = ('dispensacion', 'elaboracion', 'envasado', 'acondicionamiento')
    faltantes = [r for r in roles_esperados if asignaciones.get(r) is None]
    if faltantes:
        _log = logging.getLogger('inventario.programacion')
        _log.warning(
            'auto_asignar_operarios ABORTÓ prod=%s fecha=%s · roles sin candidato: %s · '
            'producción NO modificada (estado previo preservado)',
            produccion_id, fecha_iso, faltantes,
        )
        return None

    # UPDATE produccion_programada · valores ABSOLUTOS (ya garantizamos
    # que los 4 están). Si alguien pasó valores previos para "preservar",
    # esta función reemplaza todo el set de operarios atómicamente.
    c.execute("""
        UPDATE produccion_programada SET
          operario_dispensacion_id = ?,
          operario_elaboracion_id = ?,
          operario_envasado_id = ?,
          operario_acondicionamiento_id = ?
        WHERE id = ?
    """, (asignaciones['dispensacion'],
          asignaciones['elaboracion'],
          asignaciones['envasado'],
          asignaciones['acondicionamiento'],
          produccion_id))
    return asignaciones


def _seleccionar_area_optima(c, lote_kg, fecha_iso, excluir_sucias=True,
                               excluir_produccion_id=None):
    """IA: elige el ÁREA con tanque más chico que aguante el lote (eficiencia).

    Sebastián 1-may-2026: "si producción 200 kilos debe decir producción
    donde está marmita 250 litros".

    Sebastián 19-may-2026 · BUG-12 audit Planta PERFECTA:
    `excluir_produccion_id` evita que el query de areas_ocupadas_hoy
    cuente como ocupada la propia producción que se está re-asignando
    (antes: re-asignar una prod que ya tenía area_id la dejaba sin
    alternativa porque su propia área aparecía como "ocupada").

    Returns: {area_codigo, area_id, tanque_codigo, capacidad_litros, score, razon}
              o None si nada candidato (incluye lote_kg<=0).
    """
    # Guard: lote inválido → no se puede seleccionar área (Sebastián test)
    try:
        lote_kg = float(lote_kg)
    except (TypeError, ValueError):
        return None
    if lote_kg <= 0:
        return None
    capacidad_min = lote_kg * 1.2  # 20% headroom

    # Tanques candidatos (suficiente capacidad)
    tanques = c.execute("""
        SELECT t.id, t.codigo, t.nombre, t.area_codigo, t.capacidad_litros,
               a.id as area_id, a.estado, a.requiere_limpieza_profunda,
               a.puede_producir
        FROM equipos_planta t
          LEFT JOIN areas_planta a ON a.codigo = t.area_codigo
        WHERE t.activo = 1
          AND t.tipo IN ('tanque','marmita','olla')
          AND t.capacidad_litros IS NOT NULL
          AND t.capacidad_litros >= ?
          AND COALESCE(a.activo, 1) = 1
          AND COALESCE(a.puede_producir, 0) = 1
        ORDER BY t.capacidad_litros ASC
    """, (capacidad_min,)).fetchall()

    # Producciones ya programadas en otras áreas ese día.
    # BUG-12 fix: excluir la prod que se está re-asignando del cómputo.
    areas_ocupadas_hoy = set()
    try:
        if excluir_produccion_id is not None:
            rows = c.execute("""
                SELECT DISTINCT pp.area_id
                FROM produccion_programada pp
                WHERE date(pp.fecha_programada) = ?
                  AND pp.area_id IS NOT NULL
                  AND pp.id != ?
                  AND COALESCE(pp.estado, 'programado') NOT IN ('cancelado','completado')
            """, (fecha_iso, excluir_produccion_id)).fetchall()
        else:
            rows = c.execute("""
                SELECT DISTINCT pp.area_id
                FROM produccion_programada pp
                WHERE date(pp.fecha_programada) = ?
                  AND pp.area_id IS NOT NULL
                  AND COALESCE(pp.estado, 'programado') NOT IN ('cancelado','completado')
            """, (fecha_iso,)).fetchall()
        areas_ocupadas_hoy = {r[0] for r in rows}
    except Exception:
        pass

    candidatos = []
    seen_area = set()
    for t in tanques:
        if t[3] in seen_area:
            continue  # ya tomamos el tanque más chico de esta área
        seen_area.add(t[3])
        if not t[5]:  # area_id None
            continue
        estado_area = t[6] or 'libre'
        if excluir_sucias and estado_area == 'sucia':
            continue
        # Evitar misma área el mismo día (otro lote)
        if t[5] in areas_ocupadas_hoy:
            continue
        utilizacion = (lote_kg / t[4]) * 100
        score = 0
        razones = []
        if utilizacion >= 60:
            score += 60
            razones.append(f"Tanque {t[4]:.0f}L · uso {utilizacion:.0f}% (eficiente)")
        elif utilizacion >= 30:
            score += 40
            razones.append(f"Tanque {t[4]:.0f}L · uso {utilizacion:.0f}%")
        else:
            score += 20
            razones.append(f"Tanque {t[4]:.0f}L · uso {utilizacion:.0f}% (sub-utilizado)")
        if estado_area == 'libre':
            score += 30
            razones.append('Sala libre')
        elif estado_area == 'limpiando':
            score += 10
            razones.append('Sala en limpieza · disponible al terminar')

        candidatos.append({
            'area_codigo': t[3],
            'area_id': t[5],
            'tanque_codigo': t[1],
            'tanque_nombre': t[2],
            'capacidad_litros': t[4],
            'estado_area': estado_area,
            'utilizacion_pct': round(utilizacion, 1),
            'score': score,
            'razones': razones,
        })

    candidatos.sort(key=lambda x: x['score'], reverse=True)
    return candidatos[0] if candidatos else None


def _envasado_sugerido(area_codigo):
    """Mapeo FAB → ENV (Sebastián + Alejandro)."""
    if area_codigo in ('FAB1', 'PROD1'):
        return 'ENV1'
    if area_codigo in ('FAB2', 'PROD2', 'FAB3', 'PROD3', 'FAB_FLOAT'):
        return 'ENV2'
    return None


def _crear_limpieza_post_produccion(c, area_id, area_codigo, fecha_produccion,
                                       producto, lote, user='auto-ia'):
    """Sebastián 1-may-2026: 'queden limpias el mismo día'.
    Crea entrada en limpieza_profunda_calendario para HOY (la fecha de la
    producción) para que mañana el área esté disponible.
    """
    if not area_codigo:
        return None
    # Verificar si ya hay limpieza creada para esta área+fecha
    try:
        existe = c.execute("""
            SELECT id FROM limpieza_profunda_calendario
            WHERE area_codigo = ? AND date(fecha) = ?
              AND estado IN ('pendiente','asignada','en_proceso')
            LIMIT 1
        """, (area_codigo, fecha_produccion)).fetchone()
        if existe:
            return existe[0]
    except Exception:
        pass

    # Asignar operario rotando (excluye jefes · Sebastián 1-may-2026)
    operarios = c.execute("""
        SELECT id FROM operarios_planta
        WHERE COALESCE(activo,1)=1
          AND COALESCE(es_jefe_produccion,0)=0
        ORDER BY id
    """).fetchall()
    pool = [r[0] for r in operarios]
    if not pool:
        return None
    op_id = _siguiente_operario_rotando(c, 'limpieza', pool)
    _registrar_rotacion(c, 'limpieza', op_id, user)

    # Lookup nombre operario
    op_nombre = ''
    if op_id:
        row = c.execute(
            "SELECT nombre || ' ' || COALESCE(apellido,'') FROM operarios_planta WHERE id=?",
            (op_id,)
        ).fetchone()
        if row: op_nombre = (row[0] or '').strip()

    razon = f'Auto post-producción {producto} (lote {lote or "—"}) · limpieza mismo día'
    try:
        c.execute("""
            INSERT INTO limpieza_profunda_calendario
              (fecha, area_codigo, asignado_a, estado, generado_por, razon_asignacion)
            VALUES (?, ?, ?, 'asignada', ?, ?)
        """, (fecha_produccion, area_codigo, op_nombre, user, razon))
        return c.lastrowid
    except Exception as e:
        log.warning(f'Crear limpieza auto fallo: {e}')
        return None


@bp.route('/api/planta/auto-asignar-hoy', methods=['POST'])
def auto_asignar_hoy_bulk():
    """Re-corre la auto-asignación (área + operarios) para todas las
    producciones de HOY que NO sean Fijas (eos_plan / eos_b2b /
    eos_retroactivo). El cron de las 7am hace lo mismo automático,
    pero esto permite re-disparar a demanda · Alejandro lo usa cuando
    cambia algo de la planta y quiere re-armar el día.

    Sebastián 19-may-2026 (Operación Live · pieza 5).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    try:
        from blueprints.compras import ADMIN_USERS as _ADMIN, COMPRAS_ACCESS as _ACC
    except Exception:
        _ADMIN = ('sebastian', 'alejandro')
        _ACC = set()
    if user not in (set(_ADMIN) | set(_ACC)):
        return jsonify({'error': 'Solo admin / Compras'}), 403

    conn = get_db()
    c = conn.cursor()
    # SEC-FIX 27-may-2026 PM · audit round 5 · race condition · 2 admins + cron
    # podrían correr esto en paralelo asignando el mismo operario a 2 producciones.
    # Lock distribuido vía cron_locks · TTL corto porque la operación dura segundos.
    try:
        from blueprints.auto_plan_jobs import _adquirir_lock_cron, _liberar_lock_cron
        if not _adquirir_lock_cron(c, 'auto_asignar_hoy_bulk', ttl_horas=1):
            return jsonify({
                'error': 'Otro proceso (cron o admin) ya está asignando · reintentá en 30s',
                'codigo': 'LOCK_OCUPADO',
            }), 409
        _lock_owned = True
    except Exception:
        _lock_owned = False  # si lock helper no disponible, continúa sin lock
    rows = c.execute("""
        SELECT id, producto FROM produccion_programada
        WHERE date(fecha_programada) = date('now', '-5 hours')
          AND COALESCE(estado,'programado') NOT IN ('completado','cancelado')
          AND COALESCE(origen,'') NOT IN ('eos_plan','eos_b2b','eos_retroactivo')
    """).fetchall()
    asignadas = 0
    fallidas = []
    detalles = []
    for pid, prod in rows:
        try:
            res = _auto_asignar_produccion(c, pid, user)
            if res.get('ok'):
                asignadas += 1
                if res.get('cambios'):
                    detalles.append({'id': pid, 'producto': prod,
                                     'cambios': res.get('cambios', [])})
                # BUG-12 fix · 19-may-2026 audit Planta PERFECTA:
                # commitear DESPUÉS de cada producción exitosa para que la
                # siguiente vea el area_id asignado en la query de
                # areas_ocupadas_hoy. Antes el commit estaba afuera del
                # loop, todas las producciones del día veían area_id=NULL
                # entre sí y la IA podía asignar la MISMA área a varias.
                try:
                    conn.commit()
                except Exception as _e_commit:
                    logging.getLogger('programacion').warning(
                        f'commit intra-loop fallo pid={pid}: {_e_commit}')
            else:
                fallidas.append({'id': pid, 'producto': prod,
                                 'error': res.get('error')})
        except Exception as _e:
            fallidas.append({'id': pid, 'producto': prod,
                             'error': str(_e)[:120]})
    conn.commit()
    try:
        audit_log(c, usuario=user, accion='AUTO_ASIGNAR_HOY_BULK',
                  tabla='produccion_programada', registro_id='bulk',
                  despues={'asignadas': asignadas, 'fallidas': len(fallidas),
                           'total': len(rows)},
                  detalle=f'Auto-asignación bulk HOY · {asignadas}/{len(rows)} OK · '
                          f'respeta Fijo')
    except Exception as _e:
        logging.getLogger('programacion').warning(
            f'audit AUTO_ASIGNAR_HOY_BULK fallo: {_e}')
    # Liberar lock distribuido · siempre · incluso en error path
    if _lock_owned:
        try:
            _liberar_lock_cron(c, 'auto_asignar_hoy_bulk')
            conn.commit()
        except Exception:
            pass
    return jsonify({'ok': True, 'asignadas': asignadas, 'total': len(rows),
                    'fallidas': fallidas, 'detalles': detalles[:30]})


def _liberar_sols_pre_produccion(c, produccion_id, motivo='Producción cancelada'):
    """Helper · AUDITORÍA-FIX 23-may-2026 · C16
    Cuando se cancela una producción, las SOLs Pre-Producción que se crearon
    para esa producción específica deben cancelarse para evitar que Catalina
    apruebe compras para producción que ya no existe.

    Returns: lista de números de SOLs liberadas (para audit_log).
    """
    sols_liberadas = []
    try:
        rows = c.execute(
            """SELECT DISTINCT solicitud_numero
               FROM produccion_checklist
               WHERE produccion_id=? AND COALESCE(solicitud_numero,'') != ''""",
            (produccion_id,),
        ).fetchall()
        for r in rows:
            num = r[0] if r and r[0] else None
            if not num:
                continue
            try:
                c.execute(
                    """UPDATE solicitudes_compra
                       SET estado='Cancelada',
                           observaciones = COALESCE(observaciones,'') ||
                                          ' | ' || ?
                       WHERE numero=? AND estado IN ('Pendiente','Aprobada')""",
                    (motivo + f' (id={produccion_id})', num),
                )
                if c.rowcount > 0:
                    sols_liberadas.append(num)
            except Exception:
                pass
    except Exception:
        pass
    return sols_liberadas


def _auto_asignar_produccion(c, produccion_id, user='auto-ia'):
    """Pipeline completo de auto-asignación IA para una producción:
      1. Selecciona área óptima por lote_kg (tanque más chico que aguante)
      2. Asigna área de envasado correspondiente
      3. Asigna operarios rotando (4 roles, todos rotan)
      4. Logs en auto_asignacion_log
    """
    prod = c.execute("""
        SELECT id, producto, fecha_programada, lotes,
               COALESCE(cantidad_kg, 0), area_id,
               operario_dispensacion_id, operario_elaboracion_id,
               operario_envasado_id, operario_acondicionamiento_id
        FROM produccion_programada
        WHERE id = ?
    """, (produccion_id,)).fetchone()
    if not prod:
        return {'ok': False, 'error': 'Producción no existe'}

    fecha_iso = prod[2][:10] if prod[2] else None
    lote_kg = float(prod[4] or 0)
    if lote_kg <= 0:
        # Inferir desde fórmula
        try:
            row = c.execute("""
                SELECT COALESCE(lote_size_kg,0) FROM formula_headers
                WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
            """, (prod[1],)).fetchone()
            if row:
                lote_kg = float(row[0] or 0) * int(prod[3] or 1)
        except Exception:
            pass
    if lote_kg <= 0:
        return {'ok': False, 'error': 'lote_kg desconocido (revisa fórmula)'}

    resultado = {'ok': True, 'produccion_id': produccion_id, 'cambios': []}

    # 1) Área óptima (si no la tiene)
    if not prod[5]:
        # BUG-12 fix · 19-may-2026 audit Planta PERFECTA: pasar
        # excluir_produccion_id para que la propia prod no figure como
        # "área ocupada" si por algún motivo ya estaba apuntando a un
        # area_id que ahora se está descartando.
        area = _seleccionar_area_optima(c, lote_kg, fecha_iso,
                                          excluir_produccion_id=produccion_id)
        if not area:
            # Reintentar incluyendo sucias (con limpieza inmediata)
            area = _seleccionar_area_optima(c, lote_kg, fecha_iso,
                                              excluir_sucias=False,
                                              excluir_produccion_id=produccion_id)
            if not area:
                return {'ok': False, 'error': f'Sin área disponible para {lote_kg:.0f}kg ese día'}
            else:
                resultado['cambios'].append('⚠️ Asignada área sucia, requiere limpieza primero')

        env_codigo = _envasado_sugerido(area['area_codigo'])
        env_id = None
        if env_codigo:
            row = c.execute("SELECT id FROM areas_planta WHERE codigo=?", (env_codigo,)).fetchone()
            if row: env_id = row[0]

        # Persistir área producción + área envasado (col agregada migración 76)
        try:
            c.execute(
                "UPDATE produccion_programada SET area_id=?, area_envasado_id=? WHERE id=?",
                (area['area_id'], env_id, produccion_id)
            )
        except sqlite3.OperationalError:
            # Fallback si col area_envasado_id aún no migrada
            c.execute("UPDATE produccion_programada SET area_id=? WHERE id=?",
                      (area['area_id'], produccion_id))

        resultado['area'] = area
        resultado['area_envasado'] = {'codigo': env_codigo, 'id': env_id}
        resultado['cambios'].append(
            f"Área asignada: {area['area_codigo']} (tanque {area['tanque_codigo']} · "
            f"{area['capacidad_litros']:.0f}L · uso {area['utilizacion_pct']}%)"
            + (f" → envasado {env_codigo}" if env_codigo else "")
        )

    # 2) Operarios rotando
    # Sebastián 1-may-2026: detectar asignaciones malas (jefes asignados,
    # duplicados intra-producción) y forzar reasignación limpia.
    sin_ops = not (prod[6] or prod[7] or prod[8] or prod[9])
    necesita_reasignar = sin_ops
    if not sin_ops:
        ops_actuales = [o for o in (prod[6], prod[7], prod[8], prod[9]) if o]
        # Hay duplicados?
        if len(ops_actuales) != len(set(ops_actuales)):
            necesita_reasignar = True
            resultado['cambios'].append('⚠️ Reasignando: operarios duplicados detectados')
        else:
            # Hay algún jefe entre los asignados?
            jefes_asignados = c.execute(
                f"SELECT id FROM operarios_planta WHERE id IN ({','.join(['?']*len(ops_actuales))}) AND COALESCE(es_jefe_produccion,0)=1",
                ops_actuales
            ).fetchall()
            if jefes_asignados:
                necesita_reasignar = True
                resultado['cambios'].append('⚠️ Reasignando: jefe asignado como operario (Luis Enrique no rota)')
    if necesita_reasignar:
        # BUG-11 fix · 19-may-2026: ya NO NULLeamos los operarios previos
        # ANTES de llamar _auto_asignar_operarios. La función ahora hace
        # UPDATE absoluto si y solo si puede llenar los 4 roles, o aborta
        # sin tocar nada si pool no alcanza. Así si el caller falla,
        # el estado previo (jefes / duplicados) queda intacto para
        # debugging en vez de quedar parcialmente NULL.
        asigns = _auto_asignar_operarios(c, produccion_id, fecha_iso, user)
        if asigns:
            # Lookup nombres
            ids = [v for v in asigns.values() if v]
            nombres_map = {}
            if ids:
                placeholders = ','.join(['?'] * len(ids))
                rows = c.execute(
                    f"SELECT id, nombre || ' ' || COALESCE(apellido,'') FROM operarios_planta WHERE id IN ({placeholders})",
                    ids
                ).fetchall()
                nombres_map = {r[0]: (r[1] or '').strip() for r in rows}
            asigns_nom = {rol: nombres_map.get(oid, f'#{oid}') for rol, oid in asigns.items()}
            resultado['operarios'] = asigns_nom
            resultado['cambios'].append(
                'Operarios asignados (rotación): ' +
                ' · '.join(f'{k}: {v}' for k, v in asigns_nom.items())
            )
        else:
            # _auto_asignar_operarios devolvió None · pool no alcanza
            resultado['cambios'].append(
                '⚠️ No se pudo reasignar operarios (pool insuficiente o '
                'todos fijos/jefes) · operarios previos NO modificados'
            )

    # 3) Log
    try:
        import json as _json
        c.execute("""
            INSERT INTO auto_asignacion_log
              (produccion_id, ejecutado_por, area_asignada, tanque_asignado,
               area_envasado_asignada, operarios_json, score, razon)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            produccion_id, user,
            resultado.get('area', {}).get('area_codigo'),
            resultado.get('area', {}).get('tanque_codigo'),
            resultado.get('area_envasado', {}).get('codigo'),
            _json.dumps(resultado.get('operarios', {})),
            resultado.get('area', {}).get('score'),
            ' | '.join(resultado.get('cambios', [])),
        ))
    except Exception:
        pass

    return resultado


@bp.route('/api/planta/estado-salas-vivo', methods=['GET'])
def estado_salas_vivo():
    """Estado live de las áreas de producción + envasado.
    Sebastián 1-may-2026: vista en vivo de qué pasa en planta hoy.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()
    fecha_man = fecha_hoy + timedelta(days=1)

    salas = c.execute("""
        SELECT a.id, a.codigo, a.nombre, a.tipo, a.estado,
               a.requiere_limpieza_profunda, a.ultima_limpieza_profunda,
               a.puede_producir, a.puede_envasar
        FROM areas_planta a
        WHERE a.activo=1 AND a.tipo='produccion'
        ORDER BY a.orden, a.codigo
    """).fetchall()

    out = []
    for s in salas:
        sala = {
            'id': s[0], 'codigo': s[1], 'nombre': s[2],
            'estado': s[4] or 'libre',
            'requiere_limpieza_profunda': bool(s[5]),
            'ultima_limpieza_profunda': s[6],
            'puede_producir': bool(s[7]),
            'puede_envasar': bool(s[8]),
        }
        # Producción actual o próxima
        prod = c.execute("""
            SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
                   COALESCE(pp.cantidad_kg,0), pp.estado,
                   o1.nombre as op_disp, o2.nombre as op_elab,
                   o3.nombre as op_env
            FROM produccion_programada pp
              LEFT JOIN operarios_planta o1 ON o1.id = pp.operario_dispensacion_id
              LEFT JOIN operarios_planta o2 ON o2.id = pp.operario_elaboracion_id
              LEFT JOIN operarios_planta o3 ON o3.id = pp.operario_envasado_id
            WHERE pp.area_id = ?
              AND date(pp.fecha_programada) >= ?
              AND date(pp.fecha_programada) <= ?
              AND COALESCE(pp.estado, 'programado') NOT IN ('completado', 'cancelado')
            ORDER BY pp.fecha_programada ASC LIMIT 1
        """, (s[0], fecha_hoy.isoformat(), fecha_man.isoformat())).fetchall()
        if prod:
            p = prod[0]
            sala['produccion'] = {
                'id': p[0], 'producto': p[1], 'fecha': p[2][:10] if p[2] else '',
                'lotes': p[3], 'kg': p[4], 'estado': p[5],
                'op_dispensacion': p[6] or '',
                'op_elaboracion': p[7] or '',
                'op_envasado': p[8] or '',
            }
        else:
            sala['produccion'] = None
        # Limpieza pendiente
        try:
            limp = c.execute("""
                SELECT id, fecha, asignado_a, estado, razon_asignacion
                FROM limpieza_profunda_calendario
                WHERE area_codigo = ?
                  AND date(fecha) >= ?
                  AND estado IN ('pendiente','asignada','en_proceso')
                ORDER BY fecha ASC LIMIT 1
            """, (s[1], fecha_hoy.isoformat())).fetchone()
            if limp:
                sala['limpieza_pendiente'] = {
                    'id': limp[0], 'fecha': limp[1], 'asignado_a': limp[2] or '',
                    'estado': limp[3], 'razon': limp[4] or '',
                }
            else:
                sala['limpieza_pendiente'] = None
        except Exception:
            sala['limpieza_pendiente'] = None
        # Tanque más grande de la sala (info)
        try:
            tan = c.execute("""
                SELECT codigo, nombre, capacidad_litros
                FROM equipos_planta
                WHERE area_codigo=? AND activo=1
                  AND tipo IN ('tanque','marmita','olla')
                ORDER BY capacidad_litros DESC LIMIT 1
            """, (s[1],)).fetchone()
            if tan:
                sala['tanque_principal'] = {
                    'codigo': tan[0], 'nombre': tan[1], 'litros': tan[2]
                }
        except Exception:
            pass
        out.append(sala)

    return jsonify({
        'fecha_hoy': fecha_hoy.isoformat(),
        'salas': out,
        'total': len(out),
        'libres': sum(1 for s in out if s['estado'] == 'libre'),
        'ocupadas': sum(1 for s in out if s['estado'] == 'ocupada'),
        'sucias': sum(1 for s in out if s['estado'] == 'sucia'),
        'limpiando': sum(1 for s in out if s['estado'] == 'limpiando'),
    })


@bp.route('/api/planta/auto-asignar/<int:prod_id>', methods=['POST'])
def auto_asignar_endpoint(prod_id):
    """Auto-asigna área + envasado + operarios para 1 producción específica."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', 'manual')
    conn = get_db(); c = conn.cursor()
    try:
        resultado = _auto_asignar_produccion(c, prod_id, user)
        if resultado.get('ok'):
            conn.commit()
        return jsonify(resultado)
    except Exception as e:
        try: conn.rollback()
        except Exception as _r:
            logging.getLogger('programacion').debug('rollback no aplicable: %s', _r)
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/planta/auto-asignar-pendientes', methods=['POST'])
def auto_asignar_pendientes():
    """Recorre producciones próximos N días sin área/operarios y auto-asigna.

    Sebastián 1-may-2026: "haz todo automático".
    Acepta sesión o ?clave=AUTO_PLAN_CRON_KEY para uso interno.
    Body opcional: {dias: 7}
    """
    clave_cron = request.args.get('clave', '') or (request.json or {}).get('clave', '')
    clave_esperada = os.environ.get('AUTO_PLAN_CRON_KEY', '')
    es_cron = bool(clave_esperada and clave_cron == clave_esperada)
    if not es_cron and 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    user = 'cron-auto-asignar' if es_cron else session.get('compras_user', 'manual')
    d = request.json or {}
    dias = int(d.get('dias', 7))
    conn = get_db(); c = conn.cursor()
    fecha_hoy = datetime.now().date()
    fecha_max = fecha_hoy + timedelta(days=dias)

    pendientes = c.execute("""
        SELECT id FROM produccion_programada
        WHERE date(fecha_programada) >= ?
          AND date(fecha_programada) <= ?
          AND COALESCE(estado, 'programado') NOT IN ('completado', 'cancelado')
          AND (area_id IS NULL
               OR (operario_dispensacion_id IS NULL
                   AND operario_elaboracion_id IS NULL
                   AND operario_envasado_id IS NULL))
        ORDER BY fecha_programada ASC
    """, (fecha_hoy.isoformat(), fecha_max.isoformat())).fetchall()

    procesadas = []
    errores = []
    for (prod_id,) in pendientes:
        res = _auto_asignar_produccion(c, prod_id, user)
        if res.get('ok'):
            procesadas.append({'id': prod_id, 'cambios': res.get('cambios', [])})
        else:
            errores.append({'id': prod_id, 'error': res.get('error')})

    conn.commit()
    return jsonify({
        'ok': True,
        'fecha_hoy': fecha_hoy.isoformat(),
        'horizonte_dias': dias,
        'pendientes_evaluadas': len(pendientes),
        'procesadas': procesadas,
        'errores': errores,
        'mensaje': f'✅ {len(procesadas)} producciones auto-asignadas, {len(errores)} con error',
    })


# ════════════════════════════════════════════════════════════════════════
# PLANTA INTELIGENTE — FASE 2: Motor de Gates (Pre-Flight Checks)
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "programado un producto dice donde como, le dice
# inteligentemente area sucia confirmar limpieza confirmar tal y tal cosa".
# Antes de iniciar produccion, se corren N validaciones automaticas y se
# devuelve qué falta. Cada gate tiene status (ok/warn/blocker), mensaje y
# accion sugerida.

def _gate_sala_asignada(produccion, conn):
    """Gate: la producción tiene un área asignada."""
    if not produccion['area_id']:
        return {
            'gate': 'sala_asignada',
            'status': 'blocker',
            'titulo': 'Sala sin asignar',
            'mensaje': 'Esta producción no tiene un área asignada. Usa "Sugerir área" o asigna manualmente.',
            'accion': 'asignar_area',
        }
    area = conn.execute(
        "SELECT codigo, nombre, estado, puede_producir, activo FROM areas_planta WHERE id=?",
        (produccion['area_id'],)
    ).fetchone()
    if not area:
        return {'gate': 'sala_asignada', 'status': 'blocker',
                'titulo': 'Sala inválida', 'mensaje': f'area_id={produccion["area_id"]} no existe',
                'accion': 'asignar_area'}
    if not area[4]:
        return {'gate': 'sala_asignada', 'status': 'blocker',
                'titulo': 'Sala desactivada',
                'mensaje': f'{area[1]} está desactivada en areas_planta',
                'accion': 'asignar_area'}
    if not area[3]:
        return {'gate': 'sala_asignada', 'status': 'warn',
                'titulo': 'Sala no es de producción',
                'mensaje': f'{area[1]} no tiene puede_producir=1 (es área de apoyo)',
                'accion': None}
    return {
        'gate': 'sala_asignada',
        'status': 'ok',
        'titulo': f'Sala: {area[1]}',
        'mensaje': f'Estado actual: {area[2]}',
        'accion': None,
        'meta': {'area_codigo': area[0], 'area_nombre': area[1], 'estado': area[2]},
    }


def _gate_sala_libre(produccion, conn):
    """Gate: la sala no está ocupada por OTRA producción en curso."""
    if not produccion['area_id']:
        return None  # Sin sala no aplica este check
    area = conn.execute(
        "SELECT codigo, nombre, estado FROM areas_planta WHERE id=?",
        (produccion['area_id'],)
    ).fetchone()
    if not area:
        return None
    estado = area[2] or 'libre'
    if estado == 'libre':
        return {'gate': 'sala_libre', 'status': 'ok',
                'titulo': 'Sala libre', 'mensaje': 'Sin producción en curso',
                'accion': None}
    if estado == 'mantenimiento':
        return {'gate': 'sala_libre', 'status': 'blocker',
                'titulo': 'Sala en mantenimiento',
                'mensaje': f'{area[1]} marcada en mantenimiento — espera o cambia de sala',
                'accion': 'cambiar_estado'}
    if estado == 'ocupada':
        # Buscar qué producción tiene en curso
        otra = conn.execute("""
            SELECT id, producto FROM produccion_programada
            WHERE area_id=? AND id != ? AND estado IN ('en_proceso','pendiente')
              AND inicio_real_at IS NOT NULL AND fin_real_at IS NULL
            LIMIT 1
        """, (produccion['area_id'], produccion['id'])).fetchone()
        if otra:
            return {'gate': 'sala_libre', 'status': 'blocker',
                    'titulo': 'Sala ocupada',
                    'mensaje': f'Producción en curso: {otra[1]} (id {otra[0]})',
                    'accion': 'esperar_o_cambiar'}
    return {'gate': 'sala_libre', 'status': 'warn',
            'titulo': f'Estado sala: {estado}',
            'mensaje': 'Verifica el estado antes de iniciar',
            'accion': None}


def _gate_arrastre_pigmento(produccion, conn):
    """Gate: detecta arrastre de producto con pigmento al siguiente producto claro.
    Sebastian (30-abr-2026): pigmento → claro = limpieza profunda obligatoria."""
    if not produccion.get('area_id'):
        return None
    # Producto actual: ¿tiene perfil de riesgo?
    perfil_actual = conn.execute("""
        SELECT tiene_pigmento, color_descripcion, riesgo_arrastre_pct
        FROM producto_perfil_riesgo
        WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
    """, (produccion['producto'],)).fetchone()
    pigmento_actual = bool(perfil_actual[0]) if perfil_actual else False

    # Última producción en la misma sala antes de (o el mismo día que) ESTA.
    # FIX P1 audit 24-may-2026 · antes filtraba `< fecha` y solo veía días
    # previos · si había 2 producciones el mismo día con pigmento → claro
    # en la misma sala, el gate no detectaba arrastre. Ahora `<= fecha`
    # captura producciones AM y PM del mismo día. También se incluye
    # estado 'programado' para detectar conflicto antes de que la 1ª inicie.
    prev = conn.execute("""
        SELECT pp.producto, pp.fecha_programada
        FROM produccion_programada pp
        WHERE pp.area_id = ?
          AND pp.id != ?
          AND pp.fecha_programada <= COALESCE(?, date('now', '-5 hours', '+1 days'))
          AND pp.estado IN ('programado','completado','en_proceso')
        ORDER BY pp.fecha_programada DESC, pp.id DESC LIMIT 1
    """, (produccion['area_id'], produccion['id'], produccion.get('fecha_programada'))).fetchone()
    if not prev:
        return None
    perfil_prev = conn.execute("""
        SELECT tiene_pigmento, color_descripcion, riesgo_arrastre_pct
        FROM producto_perfil_riesgo
        WHERE UPPER(TRIM(producto_nombre)) = UPPER(TRIM(?))
    """, (prev[0],)).fetchone()
    if not perfil_prev:
        return None
    pigmento_prev = bool(perfil_prev[0])
    riesgo_prev = perfil_prev[2] or 0

    # Caso crítico: previo tenía pigmento Y actual no
    if pigmento_prev and not pigmento_actual and riesgo_prev >= 50:
        return {'gate': 'arrastre_pigmento', 'status': 'blocker',
                'titulo': '🎨 Riesgo de arrastre',
                'mensaje': f'Producto previo "{prev[0]}" tiene pigmento ({perfil_prev[1] or ""}). Limpieza profunda obligatoria antes de iniciar.',
                'accion': 'confirmar_limpieza',
                'meta': {'producto_previo': prev[0], 'fecha_previo': prev[1], 'riesgo_pct': riesgo_prev}}
    if pigmento_prev and not pigmento_actual and riesgo_prev >= 25:
        return {'gate': 'arrastre_pigmento', 'status': 'warn',
                'titulo': '🎨 Posible arrastre',
                'mensaje': f'Producto previo "{prev[0]}" pudo dejar residuo. Confirmar limpieza intermedia.',
                'accion': 'confirmar_limpieza'}
    return None


def _gate_sala_limpia(produccion, conn):
    """Gate: la sala tuvo limpieza profunda en los últimos 7 días.
    Sebastian (30-abr-2026): "area sucia confirmar limpieza"."""
    if not produccion['area_id']:
        return None
    area = conn.execute("""
        SELECT codigo, nombre, requiere_limpieza_profunda, ultima_limpieza_profunda
        FROM areas_planta WHERE id=?
    """, (produccion['area_id'],)).fetchone()
    if not area or not area[2]:
        # No requiere limpieza profunda → ok
        return {'gate': 'sala_limpia', 'status': 'ok',
                'titulo': 'Limpieza N/A',
                'mensaje': 'Esta área no requiere limpieza profunda registrada',
                'accion': None}
    # Buscar último evento de limpieza en area_eventos
    last_evt = conn.execute("""
        SELECT ts FROM area_eventos
        WHERE area_id=? AND tipo IN ('fin_limpieza','inicio_limpieza')
        ORDER BY ts DESC LIMIT 1
    """, (produccion['area_id'],)).fetchone()
    last_iso = (last_evt[0] if last_evt else None) or area[3]
    if not last_iso:
        return {'gate': 'sala_limpia', 'status': 'warn',
                'titulo': 'Sin registro de limpieza profunda',
                'mensaje': 'No hay evidencia de limpieza profunda — confirma antes de iniciar',
                'accion': 'confirmar_limpieza'}
    try:
        from datetime import datetime as _dt
        last_dt = _dt.fromisoformat((last_iso or '').replace('Z', '').split('.')[0].replace('T', ' ').strip())
        dias = (_dt.now() - last_dt).days
    except Exception:
        return {'gate': 'sala_limpia', 'status': 'warn',
                'titulo': 'Fecha de limpieza inválida',
                'mensaje': last_iso, 'accion': 'confirmar_limpieza'}
    if dias <= 7:
        return {'gate': 'sala_limpia', 'status': 'ok',
                'titulo': f'Limpieza hace {dias}d',
                'mensaje': f'Última limpieza profunda: {last_iso[:16]}',
                'accion': None}
    if dias <= 14:
        return {'gate': 'sala_limpia', 'status': 'warn',
                'titulo': f'Limpieza hace {dias}d (>7d)',
                'mensaje': 'Considera una limpieza profunda antes de iniciar',
                'accion': 'confirmar_limpieza'}
    return {'gate': 'sala_limpia', 'status': 'blocker',
            'titulo': f'Limpieza hace {dias}d (>14d)',
            'mensaje': 'Limpieza profunda obligatoria antes de iniciar',
            'accion': 'confirmar_limpieza'}


def _mp_reservada_por_dia_g(conn, fecha, exclude_evento_id=None):
    """FIX P0 audit Planta 24-may-2026 · MP comprometida por OTRAS
    producciones del día que aún no iniciaron (NO descontaron kardex todavía).

    Antes el pre-flight veía el stock completo aunque hubiera 2 producciones
    del mismo día compitiendo por la misma MP. La 1ª pasaba, iniciaba,
    descontaba kardex · la 2ª moría en /iniciar con "no hay stock" y dejaba
    la programación en limbo.

    Devuelve dict {material_id_upper: gramos_reservados}.
    `exclude_evento_id` excluye la producción actual para no doble-contar.
    """
    try:
        params = [fecha]
        sql_excl = ''
        if exclude_evento_id is not None:
            sql_excl = ' AND pp.id != ?'
            params.append(exclude_evento_id)
        rows = conn.execute(f"""
            SELECT UPPER(TRIM(fi.material_id)) AS mat,
                   SUM(
                     CASE
                       WHEN fi.cantidad_g_por_lote IS NOT NULL AND fi.cantidad_g_por_lote > 0
                            AND fh.lote_size_kg IS NOT NULL AND fh.lote_size_kg > 0
                       THEN fi.cantidad_g_por_lote * (COALESCE(pp.cantidad_kg, fh.lote_size_kg) / fh.lote_size_kg)
                       ELSE (COALESCE(fi.porcentaje, 0) / 100.0) * COALESCE(pp.cantidad_kg, fh.lote_size_kg, 0) * 1000
                     END
                   ) AS reservado_g
            FROM produccion_programada pp
            JOIN formula_headers fh
              ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
            JOIN formula_items fi
              ON UPPER(TRIM(fi.producto_nombre)) = UPPER(TRIM(pp.producto))
            WHERE pp.fecha_programada = ?
              AND COALESCE(pp.estado, 'programado') IN ('programado', 'en_proceso')
              AND pp.inicio_real_at IS NULL
              AND fi.material_id IS NOT NULL
              {sql_excl}
            GROUP BY UPPER(TRIM(fi.material_id))
        """, params).fetchall()
        return {(r[0] or ''): float(r[1] or 0) for r in rows}
    except Exception:
        return {}


def _gate_mp_disponibles(produccion, conn):
    """Gate: hay suficientes MP para el lote.
    Reusa la lógica de listo-producir.

    FIX P0 audit 24-may-2026 · descuenta MP reservada por otras producciones
    del mismo día (estado programado, inicio_real_at IS NULL) — antes el gate
    no veía esas reservas y la 2ª producción del día pasaba pre-flight pero
    moría en /iniciar.
    """
    producto = produccion['producto']
    lotes = produccion['lotes'] or 1
    fecha_prog = produccion.get('fecha_programada') or produccion.get('fecha')
    evento_id = produccion.get('id') or produccion.get('evento_id')
    fh = conn.execute(
        "SELECT lote_size_kg FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    if not fh:
        return {'gate': 'mp_disponibles', 'status': 'warn',
                'titulo': 'Producto sin fórmula',
                'mensaje': f'"{producto}" no tiene fórmula registrada',
                'accion': 'crear_formula'}
    items = conn.execute("""
        SELECT fi.material_id, fi.material_nombre, fi.porcentaje,
               COALESCE(fi.cantidad_g_por_lote, 0) as cant_lote_g
        FROM formula_items fi
        WHERE UPPER(TRIM(fi.producto_nombre))=UPPER(TRIM(?))
    """, (producto,)).fetchall()
    if not items:
        return {'gate': 'mp_disponibles', 'status': 'warn',
                'titulo': 'Fórmula sin items',
                'mensaje': 'La fórmula no tiene materiales registrados',
                'accion': 'completar_formula'}
    reservas_dia = _mp_reservada_por_dia_g(conn, fecha_prog, evento_id) if fecha_prog else {}
    deficit = []
    justo = []
    reservadas = []
    for mat_id, mat_nom, pct, cant_lote_g in items:
        req_g = (cant_lote_g or 0) * lotes if cant_lote_g else (pct or 0) / 100.0 * (fh[0] or 0) * 1000 * lotes
        # Audit zero-error 2-may-2026: usar helper canónico · excluye cuarentena
        # FIX 1-jun-2026 · resolver id fórmula→bodega antes del lookup (caso glucosamina)
        # · el gate bloqueaba producción que el descuento real sí podía ejecutar.
        _mid_bod = _resolver_material_bodega(conn, mat_id, mat_nom)
        disp_g_total = stock_mp_disponible(conn, _mid_bod)
        reservado_g = reservas_dia.get((mat_id or '').upper().strip(), 0.0)
        disp_g = disp_g_total - reservado_g
        if reservado_g > 0 and disp_g_total >= req_g and disp_g < req_g:
            # Hay stock total, pero está reservado por otra producción del día.
            reservadas.append({'codigo': mat_id, 'nombre': mat_nom,
                               'requerido_g': round(req_g),
                               'disponible_g': round(disp_g_total),
                               'reservado_otra_g': round(reservado_g),
                               'libre_g': round(disp_g)})
        if disp_g < req_g:
            if disp_g >= req_g * 0.5:
                justo.append({'codigo': mat_id, 'nombre': mat_nom,
                              'requerido_g': round(req_g), 'disponible_g': round(disp_g)})
            else:
                deficit.append({'codigo': mat_id, 'nombre': mat_nom,
                                'requerido_g': round(req_g), 'disponible_g': round(disp_g),
                                'faltante_g': round(req_g - disp_g)})
    if deficit:
        # Si TODO el déficit se explica por MP reservada por otra producción
        # del día → blocker "conflicto de reservas" (accionable: mover una de
        # las dos producciones a otro día o aumentar compra).
        if reservadas and not [d for d in deficit if d['codigo'].upper().strip() not in {r['codigo'].upper().strip() for r in reservadas}]:
            return {'gate': 'mp_disponibles', 'status': 'blocker',
                    'titulo': f'⚠ MP reservada por otra producción del día',
                    'mensaje': f'{len(reservadas)} MP ya comprometidas: ' + ', '.join(r['nombre'] for r in reservadas[:3]) + ' · movele la fecha a una de las dos producciones',
                    'accion': 'resolver_conflicto_reservas',
                    'meta': {'deficit': deficit, 'justo': justo, 'reservadas': reservadas}}
        return {'gate': 'mp_disponibles', 'status': 'blocker',
                'titulo': f'Faltan {len(deficit)} MP',
                'mensaje': f'{len(deficit)} materiales en déficit · ' + ', '.join(d['nombre'] for d in deficit[:3]),
                'accion': 'crear_tareas_compra',
                'meta': {'deficit': deficit, 'justo': justo, 'reservadas': reservadas}}
    if justo:
        return {'gate': 'mp_disponibles', 'status': 'warn',
                'titulo': f'{len(justo)} MP justos',
                'mensaje': f'{len(justo)} materiales con stock <100% del requerido (≥50%)',
                'accion': None,
                'meta': {'justo': justo, 'reservadas': reservadas}}
    return {'gate': 'mp_disponibles', 'status': 'ok',
            'titulo': 'MP suficientes',
            'mensaje': f'{len(items)} materiales disponibles',
            'accion': None}


def _gate_envases_listos(produccion, conn):
    """Gate: hay envases (MEE) suficientes para la presentación elegida."""
    producto = produccion['producto']
    # Buscar presentaciones del producto
    pres = conn.execute("""
        SELECT presentacion_codigo, etiqueta, envase_codigo, factor_g_por_unidad
        FROM producto_presentaciones
        WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND activo=1
    """, (producto,)).fetchall()
    if not pres:
        return {'gate': 'envases_listos', 'status': 'warn',
                'titulo': 'Sin presentaciones',
                'mensaje': f'"{producto}" no tiene presentaciones definidas (Fase 0)',
                'accion': 'definir_presentaciones'}
    # Para esta version: si hay presentación con envase_codigo, buscar stock en maestro_mee
    items_pres = []
    falta_envase = []
    for p_code, etiqueta, env_code, factor in pres:
        if not env_code:
            items_pres.append({'presentacion': etiqueta, 'envase': None,
                               'mensaje': 'Sin código de envase asignado'})
            continue
        # Buscar stock en maestro_mee
        mee = conn.execute(
            "SELECT stock_actual, nombre FROM maestro_mee WHERE codigo=?", (env_code,)
        ).fetchone()
        stock = (mee[0] if mee else 0) or 0
        nombre_mee = (mee[1] if mee else '') or env_code
        items_pres.append({'presentacion': etiqueta, 'envase': env_code,
                           'envase_nombre': nombre_mee, 'stock': stock})
        if stock <= 0:
            falta_envase.append(f'{etiqueta} ({env_code})')
    if falta_envase:
        return {'gate': 'envases_listos', 'status': 'warn',
                'titulo': f'Stock 0 en {len(falta_envase)} envase(s)',
                'mensaje': ', '.join(falta_envase[:3]),
                'accion': 'verificar_stock_mee',
                'meta': {'items': items_pres}}
    return {'gate': 'envases_listos', 'status': 'ok',
            'titulo': f'{len(pres)} presentación(es) OK',
            'mensaje': 'Envases con stock disponible',
            'accion': None,
            'meta': {'items': items_pres}}


def _gate_operarios(produccion, conn):
    """Gate: hay operarios asignados (al menos elaboración)."""
    asignados = []
    if produccion['operario_dispensacion_id']:
        asignados.append('dispensación')
    if produccion['operario_elaboracion_id']:
        asignados.append('elaboración')
    if produccion['operario_envasado_id']:
        asignados.append('envasado')
    if produccion['operario_acondicionamiento_id']:
        asignados.append('acondicionamiento')
    if not asignados:
        return {'gate': 'operarios', 'status': 'blocker',
                'titulo': 'Sin operarios asignados',
                'mensaje': 'Asigna al menos un operario por fase',
                'accion': 'asignar_operarios'}
    if 'elaboración' not in asignados:
        return {'gate': 'operarios', 'status': 'warn',
                'titulo': f'{len(asignados)}/4 fases con operario',
                'mensaje': 'Falta operario para elaboración',
                'accion': 'asignar_operarios'}
    return {'gate': 'operarios', 'status': 'ok',
            'titulo': f'{len(asignados)}/4 fases asignadas',
            'mensaje': ', '.join(asignados),
            'accion': None}


@bp.route('/api/planta/preflight/<int:produccion_id>', methods=['GET'])
def planta_preflight(produccion_id):
    """Motor de gates pre-flight.

    Sebastian (30-abr-2026): "programado un producto dice donde como,
    le dice inteligentemente area sucia confirmar limpieza confirmar
    tal y tal cosa". Devuelve checks ordenados (blockers primero) con
    status (ok/warn/blocker), mensaje y acción sugerida.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    pp = c.execute("""
        SELECT id, producto, fecha_programada, lotes, estado, area_id,
               operario_dispensacion_id, operario_elaboracion_id,
               operario_envasado_id, operario_acondicionamiento_id,
               inicio_real_at, fin_real_at, cantidad_kg
        FROM produccion_programada WHERE id=?
    """, (produccion_id,)).fetchone()
    if not pp:
        return jsonify({'error': 'Producción no existe'}), 404
    cols = ['id', 'producto', 'fecha_programada', 'lotes', 'estado', 'area_id',
            'operario_dispensacion_id', 'operario_elaboracion_id',
            'operario_envasado_id', 'operario_acondicionamiento_id',
            'inicio_real_at', 'fin_real_at', 'cantidad_kg']
    produccion = dict(zip(cols, pp))

    gates = []
    for fn in (_gate_sala_asignada, _gate_sala_libre, _gate_arrastre_pigmento,
               _gate_sala_limpia, _gate_mp_disponibles, _gate_envases_listos,
               _gate_operarios):
        try:
            g = fn(produccion, c)
            if g:
                gates.append(g)
        except Exception as e:
            gates.append({'gate': fn.__name__, 'status': 'warn',
                          'titulo': 'Error en check', 'mensaje': str(e), 'accion': None})

    # Resumen general
    n_block = sum(1 for g in gates if g['status'] == 'blocker')
    n_warn  = sum(1 for g in gates if g['status'] == 'warn')
    n_ok    = sum(1 for g in gates if g['status'] == 'ok')
    if n_block:
        listo = False
        veredicto = f'⛔ NO PUEDE INICIAR — {n_block} bloqueante(s)'
    elif n_warn:
        listo = True
        veredicto = f'⚠ Puede iniciar con precaución — {n_warn} advertencia(s)'
    else:
        listo = True
        veredicto = '✅ Listo para iniciar'
    # Ordenar: blockers primero, luego warns, luego ok
    orden = {'blocker': 0, 'warn': 1, 'ok': 2}
    gates.sort(key=lambda g: orden.get(g['status'], 3))

    return jsonify({
        'produccion_id': produccion_id,
        'producto': produccion['producto'],
        'estado': produccion['estado'],
        'lotes': produccion['lotes'],
        'fecha_programada': produccion['fecha_programada'],
        'gates': gates,
        'resumen': {'ok': n_ok, 'warn': n_warn, 'blocker': n_block, 'total': len(gates)},
        'listo': listo,
        'veredicto': veredicto,
    })


# ════════════════════════════════════════════════════════════════════════
# PLANTA INTELIGENTE — FASE 3: Triggers automáticos
# ════════════════════════════════════════════════════════════════════════
# Sebastian + Alejandro (30-abr-2026):
#  D. Envasado iniciado → muestra micro automática (deadline 5 días)
#  F. Resultado micro ok → entra a cola de liberación 1-2/día
#  C. Scheduler limpieza profunda L-Ma-J-V rotando 9 áreas

@bp.route('/api/planta/envasado/sugerencias', methods=['GET'])
def planta_envasado_sugerencias():
    """Semi-auto envasado · pre-llena el inicio que da el jefe: áreas de envasado
    LIMPIAS + operarios sugeridos (crew de envasado). El jefe confirma o cambia en
    1 clic. El estado de limpieza sale de areas_planta.estado (no se duplica · M9)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    areas = []
    try:
        for r in c.execute(
            "SELECT codigo, nombre, COALESCE(estado,'') FROM areas_planta "
            "WHERE COALESCE(activo,1)=1 AND tipo='envasado' "
            "ORDER BY CASE COALESCE(estado,'') WHEN 'libre' THEN 0 ELSE 1 END, nombre"
        ).fetchall():
            areas.append({'codigo': r[0], 'nombre': r[1], 'estado': r[2],
                          'limpia': r[2] == 'libre'})
    except Exception:
        areas = []
    # Operarios aptos para envasado: rol envasado/todero · no fijos en dispensación · no jefe.
    operarios = []
    try:
        for r in c.execute(
            "SELECT nombre, COALESCE(apellido,''), COALESCE(rol_predeterminado,'') "
            "FROM operarios_planta WHERE COALESCE(activo,1)=1 "
            "AND COALESCE(fija_en_dispensacion,0)=0 AND COALESCE(es_jefe_produccion,0)=0 "
            "ORDER BY CASE COALESCE(rol_predeterminado,'') WHEN 'envasado' THEN 0 "
            "WHEN 'todero' THEN 1 ELSE 2 END, nombre"
        ).fetchall():
            operarios.append({'nombre': (r[0] + ' ' + r[1]).strip(), 'rol': r[2]})
    except Exception:
        operarios = []
    return jsonify({
        'ok': True, 'areas': areas, 'operarios': operarios,
        'area_sugerida': next((a['codigo'] for a in areas if a['limpia']), ''),
        'operario_sugerido': operarios[0]['nombre'] if operarios else '',
    })


@bp.route('/api/planta/envasado/iniciar', methods=['POST'])
def planta_envasado_iniciar():
    """Jefe de producción da el clock de inicio de envasado (semi-auto). Asigna
    operario + área (pre-sugeridos por /sugerencias). El área debe estar LIMPIA
    (gate avisar+override · M5). Triggers automáticos:
       1. Crea registro produccion_envasado (con operario_asignado + area_codigo)
       2. Marca el área como ocupada
       3. Crea muestra micro pendiente con deadline = ahora + 5 días
       4. Crea entrada en cola_liberacion estado='esperando_micro'

    Body: {produccion_id*, lote*, operario?, area_codigo?, override_area?,
           presentacion_id?, presentacion_etiqueta?, unidades_planeadas?, envase_codigo?}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    prod_id = d.get('produccion_id')
    lote = (d.get('lote') or '').strip()
    if not prod_id or not lote:
        return jsonify({'error': 'produccion_id y lote requeridos'}), 400
    conn = get_db(); c = conn.cursor()
    pp = c.execute(
        "SELECT producto FROM produccion_programada WHERE id=?", (prod_id,)
    ).fetchone()
    if not pp:
        return jsonify({'error': 'Producción no existe'}), 404
    producto = pp[0]
    pres_etiqueta = (d.get('presentacion_etiqueta') or '').strip()
    pres_id = d.get('presentacion_id')
    if pres_id and not pres_etiqueta:
        pr = c.execute(
            "SELECT etiqueta, envase_codigo FROM producto_presentaciones WHERE id=?",
            (pres_id,)
        ).fetchone()
        if pr:
            pres_etiqueta = pr[0]
            if not d.get('envase_codigo'):
                d['envase_codigo'] = pr[1]
    # 0) Semi-auto · asignación del jefe: operario + área. Gate de limpieza · el
    #    área de envasado debe estar LIMPIA (areas_planta.estado='libre'). Patrón
    #    avisar+override (no bloqueo duro · M5) · el override queda en el registro.
    operario = (d.get('operario') or '').strip()
    area_codigo = (d.get('area_codigo') or '').strip()
    if area_codigo:
        ar = c.execute(
            "SELECT COALESCE(estado,'') FROM areas_planta WHERE codigo=? AND COALESCE(activo,1)=1",
            (area_codigo,)).fetchone()
        if ar and ar[0] != 'libre' and not bool(d.get('override_area')):
            return jsonify({
                'warning': f'El área {area_codigo} NO está limpia (estado: {ar[0] or "?"}). '
                           f'Limpiá el área (rótulo F02) o confirmá para iniciar igual.',
                'requiere_override': True, 'bloqueo': 'area_no_limpia',
            }), 409
    # 1) Crear envasado (con operario asignado + área · semi-auto)
    cur = c.execute("""
        INSERT INTO produccion_envasado
          (produccion_id, producto_nombre, lote, presentacion_id,
           presentacion_etiqueta, unidades_planeadas, envase_codigo,
           iniciado_por, operario_asignado, area_codigo, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'en_proceso')
    """, (prod_id, producto, lote, pres_id, pres_etiqueta or None,
          d.get('unidades_planeadas'),
          (d.get('envase_codigo') or None),
          user, operario, area_codigo))
    envasado_id = cur.lastrowid
    # 1b) Marcar el área como ocupada (deja de estar libre mientras se envasa).
    if area_codigo:
        c.execute("UPDATE areas_planta SET estado='ocupada' "
                  "WHERE codigo=? AND COALESCE(activo,1)=1 AND estado='libre'",
                  (area_codigo,))
    # 2) Crear muestra micro pendiente con deadline 5 días
    deadline = (datetime.now() + timedelta(days=5)).strftime('%Y-%m-%d')
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    # Insertamos un registro "marcador" (sin valor todavía) por cada microorganismo
    # estandar — pero para no inflar la BD, creamos UN solo marcador con
    # microorganismo='pendiente' que CC actualiza luego con cada análisis.
    try:
        cur2 = c.execute("""
            INSERT INTO calidad_micro_resultados
              (lote, producto_nombre, microorganismo, valor, estado, fecha_analisis,
               envasado_id, deadline_resultado)
            VALUES (?, ?, 'pendiente_recoleccion', NULL, 'observacion', ?, ?, ?)
        """, (lote, producto, fecha_hoy, envasado_id, deadline))
        muestra_id = cur2.lastrowid
        c.execute("UPDATE produccion_envasado SET muestra_micro_id=? WHERE id=?",
                  (muestra_id, envasado_id))
    except Exception as e:
        # FIX 1-jun-2026 (audit): antes usaba columna 'producto' (no existe · es
        # producto_nombre) + estado='pendiente' (viola el CHECK) → SIEMPRE fallaba y
        # el marcador de muestra micro NUNCA se creaba (se perdía el deadline 5d).
        # Ahora producto_nombre + estado='observacion' (válido · el gate lo trata como
        # NO conforme → avisa). Si aún falla (esquema viejo), no romper pero LOGUEAR.
        muestra_id = None
        log.warning('marcador micro NO creado · envasado=%s lote=%s err=%s',
                    envasado_id, lote, e)
    # 3) Cola de liberación
    c.execute("""
        INSERT INTO cola_liberacion
          (envasado_id, producto_nombre, lote, presentacion_etiqueta,
           unidades, fecha_envasado, fecha_min_liberacion, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'esperando_micro')
    """, (envasado_id, producto, lote, pres_etiqueta or None,
          d.get('unidades_planeadas'), fecha_hoy, deadline))
    audit_log(c, usuario=user, accion='INICIAR_ENVASADO', tabla='produccion_envasado',
              registro_id=envasado_id,
              despues={'producto': producto, 'lote': lote,
                       'presentacion': pres_etiqueta,
                       'unidades_planeadas': d.get('unidades_planeadas'),
                       'envase_codigo': d.get('envase_codigo')},
              detalle=f"Inició envasado lote {lote} ({producto})"
                      + (f" · {pres_etiqueta}" if pres_etiqueta else ""))
    conn.commit()
    return jsonify({
        'ok': True,
        'envasado_id': envasado_id,
        'muestra_micro_id': muestra_id,
        'deadline_resultado': deadline,
        'mensaje': f'Envasado iniciado. Muestra micro pendiente — deadline {deadline} (5 días).',
    })


@bp.route('/api/planta/envasado/<int:envasado_id>/terminar', methods=['POST'])
def planta_envasado_terminar(envasado_id):
    """Operario marca fin de envasado. Body: {unidades_envasadas, notas}.

    Sebastian (1-may-2026): "que coloquen cuanto fue y de alli mismo
    descuente automaticamente envases y demas del inventario". Al terminar
    envasado se descuentan envase/tapa/etiqueta del checklist usando la
    cantidad real envasada (proporcional al plan).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    cur = c.execute("""
        UPDATE produccion_envasado SET
          estado='terminado',
          terminado_at=datetime('now', '-5 hours'),
          terminado_por=?,
          unidades_envasadas=COALESCE(?, unidades_envasadas),
          notas=COALESCE(?, notas)
        WHERE id=? AND estado='en_proceso'
    """, (user, d.get('unidades_envasadas'), d.get('notas'), envasado_id))
    if cur.rowcount == 0:
        return jsonify({'error': 'Envasado no encontrado o ya terminado'}), 404
    # Actualizar unidades en cola_liberacion si llegan
    if d.get('unidades_envasadas') is not None:
        c.execute(
            "UPDATE cola_liberacion SET unidades=? WHERE envasado_id=?",
            (d.get('unidades_envasadas'), envasado_id)
        )

    # ── Descuento MEE proporcional al envasado real (envase/tapa/etiqueta) ──
    descontados_mees = []
    try:
        env_row = c.execute("""
            SELECT produccion_id, lote, unidades_planeadas, unidades_envasadas
            FROM produccion_envasado WHERE id = ?
        """, (envasado_id,)).fetchone()
        if env_row:
            prod_id, lote_env, ud_plan, ud_env = env_row
            ud_env = int(ud_env or 0)
            if prod_id and ud_env > 0:
                descontados_mees = _descontar_mee_envasado(
                    c, prod_id, lote_env or '', ud_env,
                    int(ud_plan or 0), user
                )
    except Exception as e:
        log.warning(f'Descuento MEE envasado fallo (envasado_id={envasado_id}): {e}')

    audit_log(c, usuario=user, accion='TERMINAR_ENVASADO', tabla='produccion_envasado',
              registro_id=envasado_id,
              despues={'unidades_envasadas': d.get('unidades_envasadas'),
                       'notas': d.get('notas'),
                       'mees_descontados': len(descontados_mees)},
              detalle=f"Terminó envasado id={envasado_id}"
                      + (f" · {d.get('unidades_envasadas')} unidades" if d.get('unidades_envasadas') else ""))
    conn.commit()
    return jsonify({
        'ok': True,
        'envasado_id': envasado_id,
        'mees_descontados': descontados_mees,
        'total_mees': len(descontados_mees),
    })


@bp.route('/api/planta/cola-liberacion', methods=['GET'])
def planta_cola_liberacion():
    """Lista la cola de liberación. Querystring:
       ?estado=esperando_micro|listo_revisar|liberado|rechazado|todas (default activas)"""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    estado = (request.args.get('estado') or '').strip()
    conn = get_db(); c = conn.cursor()
    where = []
    params = []
    if estado in ('esperando_micro', 'listo_revisar', 'liberado', 'rechazado', 'reanalisis'):
        where.append("estado=?"); params.append(estado)
    elif estado != 'todas':
        # default: solo activas (no liberadas/rechazadas)
        where.append("estado IN ('esperando_micro','listo_revisar','reanalisis')")
    sql = "SELECT * FROM cola_liberacion"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY fecha_min_liberacion ASC, id ASC LIMIT 1000"
    rows = c.execute(sql, params).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    # Marcar listos_revisar los que ya pasaron deadline (auto-promoción)
    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
    for it in items:
        if it['estado'] == 'esperando_micro' and (it.get('fecha_min_liberacion') or '') <= fecha_hoy:
            c.execute(
                "UPDATE cola_liberacion SET estado='listo_revisar' WHERE id=?",
                (it['id'],)
            )
            it['estado'] = 'listo_revisar'
            it['_auto_promovido'] = True
    conn.commit()
    return jsonify({
        'items': items,
        'total': len(items),
        'listos_hoy': sum(1 for it in items if it['estado'] == 'listo_revisar'),
        'esperando': sum(1 for it in items if it['estado'] == 'esperando_micro'),
    })


@bp.route('/api/planta/cola-liberacion/<int:item_id>/disposicion', methods=['POST'])
def planta_cola_liberacion_disposicion(item_id):
    """QC/Alejandro decide la disposición del lote tras revisar el resultado micro.

    Decisión INVIMA crítica (Resolución 2214/2021 art. 10): liberación de
    producto terminado solo por Calidad/Admin con motivo si rechaza.

    Body: {disposicion: 'aprobado'|'rechazado'|'reanalizar', notas?}"""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    if user not in (set(CALIDAD_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Calidad/Admin puede liberar lotes'}), 403
    d = request.json or {}
    disposicion = (d.get('disposicion') or '').strip()
    notas = (d.get('notas') or '').strip()
    if disposicion not in ('aprobado', 'rechazado', 'reanalizar'):
        return jsonify({'error': 'disposicion debe ser aprobado/rechazado/reanalizar'}), 400
    if disposicion == 'rechazado' and len(notas) < 10:
        return jsonify({'error': 'notas (≥10 chars) requeridas para rechazar lote'}), 400
    estado_nuevo = {
        'aprobado': 'liberado',
        'rechazado': 'rechazado',
        'reanalizar': 'reanalisis',
    }[disposicion]
    conn = get_db(); c = conn.cursor()
    # Capturar antes para audit log
    antes_row = c.execute(
        "SELECT producto_nombre, lote, estado, disposicion, unidades "
        "FROM cola_liberacion WHERE id=?", (item_id,)).fetchone()
    if not antes_row:
        return jsonify({'error': 'Item no encontrado'}), 404
    antes = dict(antes_row)
    _micro_override = False  # se marca True si se libera sin micro conforme (avisar+override)
    # Sebastián 28-may-2026 · INVIMA Res 2674/2013 · BLOQUEAR liberación si el
    # lote tiene resultado micro FUERA DE SPEC industria (no apto). QC debe
    # rechazar o reanalizar primero · no se puede liberar producto no conforme.
    if disposicion == 'aprobado':
        _lote_lib = (antes.get('lote') or '').strip()
        if _lote_lib:
            try:
                _oos = c.execute(
                    "SELECT COUNT(*), GROUP_CONCAT(DISTINCT microorganismo) "
                    "FROM calidad_micro_resultados "
                    "WHERE lote=? AND estado='fuera_industria'",
                    (_lote_lib,)
                ).fetchone()
            except Exception:
                _oos = None
            if _oos and (_oos[0] or 0) > 0:
                return jsonify({
                    'error': (f'NO SE PUEDE LIBERAR · lote {_lote_lib} tiene '
                              f'{_oos[0]} resultado(s) microbiológico(s) FUERA DE '
                              f'ESPECIFICACIÓN INVIMA'),
                    'microorganismos': _oos[1] or '',
                    'hint': ('El producto no es apto. Debe RECHAZAR el lote o '
                             'REANALIZAR antes de liberar. INVIMA Res 2674/2013.'),
                    'bloqueo': 'micro_fuera_industria',
                }), 409
            # FIX 1-jun-2026 · INVIMA Res 2674/2013 · además del OOS (arriba), exigir
            # micro CONFORME (ok/fuera_meta). Si no hay → AVISAR + override explícito
            # (decisión Sebastián 1-jun: avisar+override, no bloqueo duro). El override
            # queda en audit (quién liberó sin micro conforme).
            try:
                _conf = c.execute(
                    "SELECT COUNT(*) FROM calidad_micro_resultados "
                    "WHERE lote=? AND estado IN ('ok','fuera_meta')", (_lote_lib,)
                ).fetchone()
            except Exception:
                _conf = None
            if not (_conf and (_conf[0] or 0) > 0):
                if not bool(d.get('override_micro')):
                    return jsonify({
                        'warning': (f'El lote {_lote_lib} NO tiene un resultado micro '
                                    f'CONFORME (ok/fuera_meta) registrado. Si lo liberás '
                                    f'igual, queda bajo tu responsabilidad y se registra '
                                    f'en auditoría.'),
                        'requiere_override': True,
                        'hint': 'Registrá el resultado micro primero, o confirmá para liberar igual.',
                        'bloqueo': 'micro_sin_conforme',
                    }), 409
                _micro_override = True
    c.execute("""
        UPDATE cola_liberacion SET
          disposicion=?, estado=?, aprobado_por=?, aprobado_at=datetime('now', '-5 hours'),
          notas=COALESCE(?, notas)
        WHERE id=?
    """, (disposicion, estado_nuevo, user, notas or None, item_id))
    accion = {
        'aprobado': 'LIBERAR_LOTE_PT',
        'rechazado': 'RECHAZAR_LOTE_PT',
        'reanalizar': 'REANALIZAR_LOTE_PT',
    }[disposicion]
    audit_log(c, usuario=user, accion=accion, tabla='cola_liberacion',
              registro_id=item_id, antes=antes,
              despues={'disposicion': disposicion, 'estado': estado_nuevo,
                       'aprobado_por': user, 'notas': notas,
                       'micro_override': _micro_override},
              detalle=f"{accion} lote {antes.get('lote','—')} ({antes.get('producto_nombre','')})"
                      + (' · ⚠ LIBERADO SIN MICRO CONFORME (override)' if _micro_override else '')
                      + (f" · {notas}" if notas else ""))
    conn.commit()
    return jsonify({'ok': True, 'estado': estado_nuevo})


@bp.route('/api/planta/limpieza-profunda/calendario', methods=['GET'])
def planta_limpieza_calendario():
    """Lista programación de limpieza profunda. Querystring: ?dias=20 (default)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 20))
    except (TypeError, ValueError):
        dias = 20
    fecha_desde = datetime.now().strftime('%Y-%m-%d')
    fecha_hasta = (datetime.now() + timedelta(days=dias)).strftime('%Y-%m-%d')
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT lpc.*, ap.nombre as area_nombre
        FROM limpieza_profunda_calendario lpc
        LEFT JOIN areas_planta ap ON ap.codigo = lpc.area_codigo
        WHERE lpc.fecha BETWEEN ? AND ?
        ORDER BY lpc.fecha ASC, lpc.area_codigo
    """, (fecha_desde, fecha_hasta)).fetchall()
    cols = [d[0] for d in c.description]
    items = [dict(zip(cols, r)) for r in rows]
    # Cobertura de áreas: cuántas áreas tuvieron limpieza en los últimos 14d
    cobertura = c.execute("""
        SELECT area_codigo, MAX(fecha) as ultima
        FROM limpieza_profunda_calendario
        WHERE estado='completada' AND fecha >= date('now', '-5 hours', '-14 days')
        GROUP BY area_codigo
    """).fetchall()
    return jsonify({
        'items': items,
        'rango': {'desde': fecha_desde, 'hasta': fecha_hasta, 'dias': dias},
        'cobertura_14d': {r[0]: r[1] for r in cobertura},
    })


@bp.route('/api/planta/limpieza-profunda/generar', methods=['POST'])
def planta_limpieza_generar():
    """Genera cronograma rotativo L-Ma-J-V para los próximos N días.
    Brief K (Alejandro): rotar L-Ma-J-V cubriendo las 9 áreas, idealmente
    el área que se produjo el día anterior.

    Body: {dias?: int (default 20), reset?: bool (borrar pendientes y regenerar)}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    try:
        dias = int(d.get('dias', 20))
    except (TypeError, ValueError):
        dias = 20
    conn = get_db(); c = conn.cursor()

    if d.get('reset'):
        c.execute(
            "DELETE FROM limpieza_profunda_calendario WHERE estado='programada' AND fecha >= date('now', '-5 hours')"
        )

    # Obtener orden de rotación según última limpieza por área
    ultimas = c.execute("""
        SELECT codigo, ultima_limpieza_profunda
        FROM areas_planta
        WHERE codigo IN ({})
    """.format(','.join('?' for _ in _AREAS_LIMPIEZA_PROFUNDA)),
        _AREAS_LIMPIEZA_PROFUNDA
    ).fetchall()
    # Mapear código → días desde última limpieza (-1 si nunca)
    dias_por_area = {}
    for cod, ult in ultimas:
        if ult:
            try:
                last = datetime.fromisoformat((ult or '').split('.')[0].replace('T', ' '))
                dias_por_area[cod] = (datetime.now() - last).days
            except Exception:
                dias_por_area[cod] = 999
        else:
            dias_por_area[cod] = 999
    # Cola priorizada: las que llevan más sin limpiar primero
    cola = sorted(_AREAS_LIMPIEZA_PROFUNDA, key=lambda x: dias_por_area.get(x, 999), reverse=True)

    creadas = []
    skipped = []
    today = datetime.now().date()
    for offset in range(dias + 1):
        fecha = today + timedelta(days=offset)
        # Limpieza profunda solo L-Ma-J-V (excluye Mié=2, Sáb=5, Dom=6)
        if fecha.weekday() in (2, 5, 6):
            continue
        fecha_str = fecha.strftime('%Y-%m-%d')
        # Buscar área que se produjo el día anterior (preferencia)
        prev_str = (fecha - timedelta(days=1)).strftime('%Y-%m-%d')
        prev_area = c.execute("""
            SELECT ap.codigo
            FROM produccion_programada pp
            LEFT JOIN areas_planta ap ON ap.id = pp.area_id
            WHERE date(pp.fecha_programada) = ? AND ap.codigo IN ({})
            ORDER BY pp.id DESC LIMIT 1
        """.format(','.join('?' for _ in _AREAS_LIMPIEZA_PROFUNDA)),
            [prev_str, *_AREAS_LIMPIEZA_PROFUNDA]
        ).fetchone()
        if prev_area and prev_area[0] in cola:
            elegida = prev_area[0]
            razon = f'Producción ayer en {elegida}'
            cola.remove(elegida)
            cola.append(elegida)  # mover al final para rotación
        else:
            elegida = cola[0]
            razon = f'Más antigua sin limpieza ({dias_por_area.get(elegida, 999)}d)'
            cola.append(cola.pop(0))
        try:
            c.execute("""
                INSERT INTO limpieza_profunda_calendario
                  (fecha, area_codigo, estado, generado_por, razon_asignacion)
                VALUES (?, ?, 'programada', ?, ?)
            """, (fecha_str, elegida, user, razon))
            creadas.append({'fecha': fecha_str, 'area': elegida, 'razon': razon})
        except sqlite3.IntegrityError:
            skipped.append({'fecha': fecha_str, 'area': elegida, 'motivo': 'ya existe'})
    conn.commit()
    return jsonify({
        'ok': True,
        'creadas': creadas,
        'skipped': skipped,
        'total_creadas': len(creadas),
    })


@bp.route('/api/planta/limpieza-profunda/<int:item_id>/completar', methods=['POST'])
def planta_limpieza_completar(item_id):
    """Operario marca limpieza completada. Actualiza también
    areas_planta.ultima_limpieza_profunda."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT area_codigo FROM limpieza_profunda_calendario WHERE id=?", (item_id,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Item no existe'}), 404
    area_codigo = row[0]
    c.execute("""
        UPDATE limpieza_profunda_calendario SET
          estado='completada', terminado_at=datetime('now', '-5 hours'), terminado_por=?,
          notas=COALESCE(?, notas)
        WHERE id=?
    """, (user, d.get('notas'), item_id))
    c.execute(
        "UPDATE areas_planta SET ultima_limpieza_profunda=datetime('now', '-5 hours') WHERE codigo=?",
        (area_codigo,)
    )
    c.execute("""
        INSERT INTO area_eventos (area_id, tipo, usuario, nota)
        SELECT id, 'fin_limpieza', ?, ? FROM areas_planta WHERE codigo=?
    """, (user, d.get('notas') or 'Limpieza profunda programada', area_codigo))
    conn.commit()
    return jsonify({'ok': True, 'area': area_codigo})


# ════════════════════════════════════════════════════════════════════════
# PLANTA INTELIGENTE — FASE 4: Plan semanal + cascade aceptar-producción
# ════════════════════════════════════════════════════════════════════════
# Sebastian (30-abr-2026): "programación de la semana entra mira lunes
# acido hialuronico, animus le quedan 20 dias estamos sobre lo que es...
# sale que alcanzan las materias primas que recuerda no es solo para ese
# producto siempre que calcula un producto suma los consumos de todas las
# programaciones previas a esa, entonces dice 20 dias sin alerta, si se
# esta vendiendo mas deberia generar alerta a compras gerencia y alejandro
# incluso correo... entonces lo selecciona le sale con la foto, y de una
# sale señalar envases, solicitar etiquetas, armado de goteros si requiere,
# aceptar produccion se dispone para realizar, entonces automaticamente
# pasa a que el sistema decida en que area se hace y genere todo".

def _calcular_mp_requerido(producto, lotes, conn):
    """Devuelve dict {material_id: gramos_requeridos} para una producción."""
    fh = conn.execute(
        "SELECT lote_size_kg FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
        (producto,)
    ).fetchone()
    if not fh:
        return {}
    lote_kg = fh[0] or 0
    items = conn.execute("""
        SELECT material_id, porcentaje, COALESCE(cantidad_g_por_lote, 0)
        FROM formula_items
        WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))
    """, (producto,)).fetchall()
    out = {}
    for mat_id, pct, cant_lote in items:
        if cant_lote:
            out[mat_id] = (cant_lote or 0) * lotes
        else:
            out[mat_id] = (pct or 0) / 100.0 * lote_kg * 1000 * lotes
    return out


def _stock_mp(material_id, conn):
    """Stock total de MP · audit zero-error 2-may-2026.

    Wrapper sobre stock_mp_total para compatibilidad con código legacy. El
    plan_semanal usa total (incluye cuarentena) porque calcula consumos
    futuros · si llegan en QC a tiempo, contarán.
    """
    return stock_mp_total(conn, material_id)


@bp.route('/api/planta/cronograma-areas', methods=['GET'])
def planta_cronograma_areas():
    """Matriz semanal · 10 áreas × 5 días Lun-Vie con todas las fases.

    Vista que Alejandro pidió (programacion_mayo_areas.html). Auto-alimentada
    desde data existente:
      - FAB1/FYE2/FYE3 ← produccion_programada (sala PROD1/PROD2/PROD3)
      - ENV1/ENV2     ← produccion_envasado + produccion_programada (PROD4)
      - MICRO         ← calidad_micro_resultados (fecha_muestreo)
      - LIB           ← cola_liberacion (fecha_min_liberacion)
      - ACOND         ← acondicionamiento
      - ENTR          ← despachos (fecha)
      - LIMP          ← limpieza_profunda_calendario

    Query params:
      desde · YYYY-MM-DD del lunes (default: lunes de esta semana)

    Response:
      {
        rango: {desde, hasta, semana},
        days: ['Lun 04', ..., 'Vie 08'],
        areas: {
          fab1: [[{t,l,u}, ...], [], ...],   # 5 listas (Lun-Vie)
          fye2, fye3, env1, env2, micro, lib, acond, entr, limp: idem
        }
      }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from datetime import datetime as _dt, timedelta as _td

    # Calcular lunes de la semana
    desde_param = (request.args.get('desde') or '').strip()
    if desde_param:
        try:
            d0 = _dt.strptime(desde_param[:10], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'desde inválido (YYYY-MM-DD)'}), 400
    else:
        hoy = _dt.now().date()
        d0 = hoy - _td(days=hoy.weekday())  # lunes
    # Avanzar a lunes si cae en mié/dom
    while d0.weekday() != 0:
        d0 = d0 - _td(days=1)
    fechas = [d0 + _td(days=i) for i in range(5)]  # Lun-Vie
    fechas_iso = [f.isoformat() for f in fechas]
    fechas_set = set(fechas_iso)
    desde_iso = fechas_iso[0]
    hasta_iso = fechas_iso[-1]

    days_labels_es = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie']
    days_labels = [f"{days_labels_es[i]} {fechas[i].strftime('%d')}" for i in range(5)]

    # Matriz inicializada · 10 áreas × 5 días
    areas_keys = ['fab1', 'fye2', 'fye3', 'env1', 'env2',
                  'micro', 'lib', 'acond', 'entr', 'limp']
    matriz = {k: [[] for _ in range(5)] for k in areas_keys}

    def _idx(fecha_iso):
        try:
            return fechas_iso.index(fecha_iso)
        except ValueError:
            return None

    def _add(area, idx, t, label, urgente=False):
        if idx is None: return
        chip = {'t': t, 'l': label}
        if urgente:
            chip['u'] = True
        matriz[area][idx].append(chip)

    conn = get_db(); c = conn.cursor()

    # ── FAB · produccion_programada con area_id mapeado a PROD1/2/3 ─────
    # Sala → fila de Alejandro:
    #   PROD1 → fab1
    #   PROD2 → fye2 (fase fab del día)
    #   PROD3 → fye3 (fase fab del día)
    #   PROD4 → env2 (sala dedicada a envasado en HTML de Alejandro)
    #   ENV1  → env1
    SALA_TO_FILA_FAB = {'PROD1': 'fab1', 'PROD2': 'fye2', 'PROD3': 'fye3',
                         'PROD4': 'env2'}
    try:
        rows = c.execute(f"""
            SELECT pp.fecha_programada, pp.producto, ap.codigo,
                   COALESCE(pp.observaciones,''),
                   COALESCE(pp.estado, 'pendiente')
            FROM produccion_programada pp
            LEFT JOIN areas_planta ap ON ap.id = pp.area_id
            WHERE pp.fecha_programada BETWEEN ? AND ?
              AND LOWER(COALESCE(pp.estado,'')) NOT IN ('cancelado')
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, producto, sala, obs, estado in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            fila = SALA_TO_FILA_FAB.get(sala or '', 'fab1')
            urgente = ('urg' in (obs or '').lower() or
                        'crítico' in (obs or '').lower() or
                        '⚡' in (obs or ''))
            _add(fila, idx, 'fab', producto, urgente)
    except Exception:
        pass

    # ── ENV · produccion_envasado (iniciado_at o terminado_at en rango) ─
    try:
        rows = c.execute(f"""
            SELECT date(COALESCE(iniciado_at, terminado_at, fecha_creacion)) as fecha,
                   producto_nombre, lote,
                   COALESCE(presentacion_etiqueta, ''),
                   COALESCE(estado, '')
            FROM produccion_envasado
            WHERE date(COALESCE(iniciado_at, terminado_at, fecha_creacion))
                  BETWEEN ? AND ?
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, producto, lote, presentacion, estado in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            label = f"{producto}\n{presentacion}" if presentacion else (producto or '')
            # Sin info de sala específica · van a env1 por default
            _add('env1', idx, 'env', label or 'Envasado')
    except Exception:
        pass

    # ── MICRO · calidad_micro_resultados (fecha_muestreo o fecha_analisis) ─
    try:
        rows = c.execute(f"""
            SELECT COALESCE(fecha_muestreo, fecha_analisis) as fecha,
                   producto_nombre, lote, deadline_resultado
            FROM calidad_micro_resultados
            WHERE COALESCE(fecha_muestreo, fecha_analisis) BETWEEN ? AND ?
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, producto, lote, deadline in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            label = (producto or 'Muestra MICRO')
            if deadline:
                label += f"\n→ libera {deadline[5:10].replace('-','/')}"
            _add('micro', idx, 'micro', label)
    except Exception:
        pass

    # ── LIB · cola_liberacion (fecha_min_liberacion en rango) ───────────
    try:
        rows = c.execute(f"""
            SELECT fecha_min_liberacion, producto_nombre, lote, estado
            FROM cola_liberacion
            WHERE fecha_min_liberacion BETWEEN ? AND ?
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, producto, lote, estado in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            label = f"{producto}\n({lote or 'sin-lote'})" if producto else f'Lote {lote}'
            _add('lib', idx, 'lib', label)
    except Exception:
        pass

    # ── ACOND · acondicionamiento (creado_en o fecha_inicio) ───────────
    try:
        rows = c.execute(f"""
            SELECT date(COALESCE(creado_en, datetime('now', '-5 hours'))) as fecha,
                   producto, lote, COALESCE(estado, '')
            FROM acondicionamiento
            WHERE date(COALESCE(creado_en, datetime('now', '-5 hours')))
                  BETWEEN ? AND ?
              AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, producto, lote, estado in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            _add('acond', idx, 'acond', producto or f'Lote {lote}')
    except Exception:
        pass

    # ── ENTR · despachos (fecha) ────────────────────────────────────────
    try:
        rows = c.execute(f"""
            SELECT date(d.fecha) as fecha, COALESCE(cl.nombre,'') as cliente,
                   COUNT(di.id) as items
            FROM despachos d
              LEFT JOIN clientes cl ON cl.id = d.cliente_id
              LEFT JOIN despachos_items di ON di.numero_despacho = d.numero
            WHERE date(d.fecha) BETWEEN ? AND ?
            GROUP BY d.numero
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, cliente, n_items in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            label = f"Entrega {cliente}" + (f"\n{n_items} items" if n_items else '')
            _add('entr', idx, 'entr', label.strip() or 'Entrega')
    except Exception:
        pass

    # ── LIMP · limpieza_profunda_calendario ─────────────────────────────
    try:
        rows = c.execute(f"""
            SELECT lpc.fecha, lpc.area_codigo,
                   COALESCE(ap.nombre, lpc.area_codigo) as nombre,
                   lpc.estado
            FROM limpieza_profunda_calendario lpc
            LEFT JOIN areas_planta ap ON ap.codigo = lpc.area_codigo
            WHERE lpc.fecha BETWEEN ? AND ?
              AND LOWER(COALESCE(lpc.estado,'')) NOT IN ('cancelada')
        """, (desde_iso, hasta_iso)).fetchall()
        for fecha, area_cod, nombre, estado in rows:
            idx = _idx(fecha)
            if idx is None:
                continue
            _add('limp', idx, 'limp', nombre or area_cod or 'Limpieza')
    except Exception:
        pass

    return jsonify({
        'rango': {
            'desde': desde_iso, 'hasta': hasta_iso,
            'semana': f"{fechas[0].strftime('%d')}–{fechas[-1].strftime('%d')} {fechas[0].strftime('%b')} {fechas[0].year}"
        },
        'days': days_labels,
        'areas': matriz,
    })


@bp.route('/api/planta/cronograma-comparar-alejandro', methods=['GET'])
def planta_cronograma_comparar_alejandro():
    """Compara el cronograma de fabricación que mandó Alejandro vs lo que
    está realmente en produccion_programada (Calendar sync).

    Devuelve 3 secciones:
      - matches: producto+fecha que coinciden en ambos
      - en_alejandro_no_calendar: Alejandro programó pero Calendar NO tiene
      - en_calendar_no_alejandro: Calendar tiene pero Alejandro NO mencionó

    Esto le ayuda a Sebastián entender qué va a pasar:
      → Si "en_alejandro_no_calendar" tiene mucho → falta cargar producciones
      → Si "en_calendar_no_alejandro" tiene mucho → Calendar tiene cosas
        que Alejandro no contempló (revisar si están bien)
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from datetime import datetime as _dt
    from templates_py.cronograma_alejandro_data import (
        ALEJANDRO_FAB_MAYO_2026, matchea
    )
    conn = get_db(); c = conn.cursor()

    # Producciones programadas en Calendar para mayo 2026
    rows = c.execute("""
        SELECT pp.fecha_programada, pp.producto, ap.codigo as area_cod,
               COALESCE(pp.estado, 'pendiente'), pp.id,
               COALESCE(pp.observaciones, '')
        FROM produccion_programada pp
        LEFT JOIN areas_planta ap ON ap.id = pp.area_id
        WHERE pp.fecha_programada BETWEEN '2026-05-01' AND '2026-05-31'
          AND LOWER(COALESCE(pp.estado, '')) NOT IN ('cancelado')
    """).fetchall()
    calendar_items = []
    for fecha, producto, area, estado, pid, obs in rows:
        calendar_items.append({
            'fecha': fecha, 'producto': producto or '',
            'area': area or '', 'estado': estado, 'id': pid,
            'observaciones': obs or '',
        })

    # Cross-check
    matches = []
    en_alejandro_no_calendar = []
    en_calendar_no_alejandro = list(calendar_items)  # parte como todos, voy quitando los que matchean

    for fecha_a, producto_a, area_a, urgente in ALEJANDRO_FAB_MAYO_2026:
        # Buscar match en calendar (mismo día + producto similar)
        match = None
        for cal in calendar_items:
            if cal['fecha'] == fecha_a and matchea(cal['producto'], producto_a):
                match = cal
                break
        if match:
            # Mismo día y producto matchea
            area_match = (match['area'] == area_a)
            matches.append({
                'fecha': fecha_a,
                'producto_alejandro': producto_a,
                'producto_calendar': match['producto'],
                'area_alejandro': area_a,
                'area_calendar': match['area'],
                'area_match': area_match,
                'urgente_alejandro': urgente,
                'estado': match['estado'],
                'id_calendar': match['id'],
            })
            # Quitar de "en_calendar_no_alejandro"
            en_calendar_no_alejandro = [
                x for x in en_calendar_no_alejandro if x['id'] != match['id']
            ]
        else:
            # Buscar match en cualquier fecha (para detectar cambio de fecha)
            match_otra_fecha = None
            for cal in calendar_items:
                if matchea(cal['producto'], producto_a):
                    match_otra_fecha = cal
                    break
            en_alejandro_no_calendar.append({
                'fecha': fecha_a,
                'producto': producto_a,
                'area': area_a,
                'urgente': urgente,
                'match_otra_fecha': (match_otra_fecha['fecha']
                                      if match_otra_fecha else None),
                'producto_calendar_otra_fecha': (
                    match_otra_fecha['producto'] if match_otra_fecha else None
                ),
            })

    return jsonify({
        'rango': {'desde': '2026-05-01', 'hasta': '2026-05-31'},
        'resumen': {
            'total_alejandro': len(ALEJANDRO_FAB_MAYO_2026),
            'total_calendar': len(calendar_items),
            'matches_completos': sum(1 for m in matches if m.get('area_match')),
            'matches_fecha_distinto_area': sum(1 for m in matches if not m.get('area_match')),
            'falta_en_calendar': len(en_alejandro_no_calendar),
            'extra_en_calendar': len(en_calendar_no_alejandro),
        },
        'matches': matches,
        'en_alejandro_no_calendar': en_alejandro_no_calendar,
        'en_calendar_no_alejandro': en_calendar_no_alejandro,
    })


# ─────────────────────────────────────────────────────────────────────────
#  ASIGNAR ÁREAS · UI para que Alejandro asigne sala a cada producción
# ─────────────────────────────────────────────────────────────────────────
# Sebastián 2-may-2026: dos vías complementarias para que el plan refleje
# la organización que pide Alejandro:
#   1) Convención [CODIGO] al inicio del titulo en Calendar (parser en _sync).
#   2) UI manual: pantalla con listado + dropdown + auto-sugerir + confirmar.
# Estos endpoints sirven a la opción (2). Los cambios se persisten en
# produccion_programada.area_id (misma columna que los demás flujos).

def _sugerir_area_para_producto(c, producto, lote_kg):
    """Sugiere area_id para una producción según producto + tamaño de lote.

    Reglas (basadas en project_planta_crew_areas.md):
      • Si la fórmula contiene alcohol (palabras 'alcohol', 'etanol', 'isopropi') → PROD1 (especial='alcoholes')
      • Si lote_kg ≤ 100 → PROD2 (marmita 100ml)
      • Si lote_kg ≤ 250 → PROD3 (marmita 250ml)
      • Si lote_kg > 250 → PROD1 (gran capacidad, no especial)
      • Default → PROD1
    """
    try:
        # Cargar areas y formulas
        areas = {r[1]: r[0] for r in c.execute(
            "SELECT id, codigo FROM areas_planta WHERE activo=1"
        ).fetchall()}
        if not areas:
            return None
        # Detectar alcoholes en la fórmula
        usa_alcohol = False
        try:
            mps = c.execute("""
                SELECT LOWER(COALESCE(m.descripcion,''))
                  FROM formulas_v2 f
                  LEFT JOIN maestro_mp m ON m.id = f.mp_id
                 WHERE f.producto_nombre = ?
            """, (producto,)).fetchall()
            for r in mps:
                desc = r[0] or ''
                if any(kw in desc for kw in ('alcohol', 'etanol', 'isopropi')):
                    usa_alcohol = True
                    break
        except Exception:
            pass
        if usa_alcohol and 'PROD1' in areas:
            return areas['PROD1']
        try:
            lk = float(lote_kg or 0)
        except Exception:
            lk = 0
        if lk and lk <= 100 and 'PROD2' in areas:
            return areas['PROD2']
        if lk and lk <= 250 and 'PROD3' in areas:
            return areas['PROD3']
        if 'PROD1' in areas:
            return areas['PROD1']
        # Fallback: cualquier area que pueda producir
        any_prod = c.execute(
            "SELECT id FROM areas_planta WHERE puede_producir=1 AND activo=1 LIMIT 1"
        ).fetchone()
        return any_prod[0] if any_prod else None
    except Exception:
        return None


@bp.route('/api/planta/asignar-areas', methods=['GET'])
def planta_asignar_areas_listar():
    """Lista producciones de los próximos N días con su área actual + sugerida.

    Query params:
        dias (int, default 30): horizonte hacia adelante
        solo_sin_area (bool, default false): solo las que no tienen area asignada

    Devuelve para cada producción: id, producto, fecha, lotes, cantidad_kg,
    estado, area_id_actual + nombre, area_sugerida_id + codigo, lista de
    áreas disponibles para dropdown.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 30))
    except (TypeError, ValueError):
        dias = 30
    dias = max(1, min(dias, 180))
    solo_sin = request.args.get('solo_sin_area', 'false').lower() in ('1', 'true', 'yes')

    conn = get_db(); c = conn.cursor()
    # Cargar áreas activas (para dropdown)
    areas_disp = []
    for r in c.execute(
        """SELECT id, codigo, nombre, puede_producir, puede_envasar,
                  marmita_ml, especial
             FROM areas_planta WHERE activo=1 ORDER BY orden, id"""
    ).fetchall():
        areas_disp.append({
            'id': r[0], 'codigo': r[1], 'nombre': r[2],
            'puede_producir': bool(r[3]), 'puede_envasar': bool(r[4]),
            'marmita_ml': r[5], 'especial': r[6],
        })

    # Cargar producciones del horizonte (no completadas, no canceladas)
    where = """
        WHERE pp.fecha_programada BETWEEN date('now', '-5 hours') AND date('now', '-5 hours', ?)
          AND LOWER(COALESCE(pp.estado,'')) NOT IN ('completado','cancelado')
    """
    params = [f'+{dias} day']
    if solo_sin:
        where += " AND pp.area_id IS NULL"

    rows = c.execute(f"""
        SELECT pp.id, pp.producto, pp.fecha_programada,
               COALESCE(pp.lotes,1), COALESCE(pp.cantidad_kg, 0),
               COALESCE(pp.estado,'programado'),
               pp.area_id, ap.codigo, ap.nombre,
               COALESCE(pp.observaciones,''),
               COALESCE(pp.origen,'manual')
          FROM produccion_programada pp
          LEFT JOIN areas_planta ap ON ap.id = pp.area_id
          {where}
          ORDER BY pp.fecha_programada, pp.producto
    """, params).fetchall()

    # Cache lote_size_kg por producto (para sugerir)
    formulas = _get_formulas(conn)
    items = []
    for r in rows:
        (pid, prod, fecha, lotes, cant_kg, estado, area_id_act,
         area_cod, area_nom, obs, origen) = r
        lote_kg = formulas.get(prod, {}).get('lote_size_kg', 0) or 0
        sug_id = _sugerir_area_para_producto(c, prod, lote_kg)
        sug_cod = next((a['codigo'] for a in areas_disp if a['id'] == sug_id), None)
        items.append({
            'id': pid,
            'producto': prod or '',
            'fecha': fecha,
            'lotes': lotes,
            'cantidad_kg': float(cant_kg or 0),
            'estado': estado,
            'origen': origen,
            'observaciones': obs,
            'area_id_actual': area_id_act,
            'area_codigo_actual': area_cod,
            'area_nombre_actual': area_nom,
            'area_sugerida_id': sug_id,
            'area_sugerida_codigo': sug_cod,
            'lote_size_kg': float(lote_kg or 0),
        })

    return jsonify({
        'horizonte_dias': dias,
        'solo_sin_area': solo_sin,
        'total': len(items),
        'sin_area': sum(1 for x in items if x['area_id_actual'] is None),
        'areas_disponibles': areas_disp,
        'producciones': items,
    })


@bp.route('/api/planta/asignar-areas', methods=['POST'])
def planta_asignar_areas_bulk():
    """Asigna área a múltiples producciones en una sola llamada.

    Body JSON:
        asignaciones: [{id: int, area_id: int|null}, ...]

    Valida cada asignación. Devuelve dict con éxitos, errores y warnings
    (sala-ocupada en mismo día). NO falla la operación completa si una
    asignación individual no es válida — reporta y sigue con el resto.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json(force=True, silent=True) or {}
    asignaciones = data.get('asignaciones', [])
    if not isinstance(asignaciones, list) or not asignaciones:
        return jsonify({'error': 'asignaciones debe ser lista no vacía'}), 400
    if len(asignaciones) > 200:
        return jsonify({'error': 'demasiadas asignaciones (max 200)'}), 400

    conn = get_db(); c = conn.cursor()
    user = session.get('compras_user', 'desconocido')
    ok_ids, errores, warnings = [], [], []

    # Cache áreas para validar y reportar conflictos
    areas_map = {}
    for r in c.execute(
        "SELECT id, codigo, nombre, puede_producir FROM areas_planta WHERE activo=1"
    ).fetchall():
        areas_map[r[0]] = {'codigo': r[1], 'nombre': r[2], 'puede_producir': bool(r[3])}

    for asg in asignaciones:
        try:
            pid = int(asg.get('id'))
        except (TypeError, ValueError):
            errores.append({'id': asg.get('id'), 'error': 'id inválido'})
            continue
        new_area = asg.get('area_id')
        if new_area is not None:
            try:
                new_area = int(new_area)
            except (TypeError, ValueError):
                errores.append({'id': pid, 'error': 'area_id inválido'})
                continue
            if new_area not in areas_map:
                errores.append({'id': pid, 'error': 'área no existe o inactiva'})
                continue

        pp = c.execute(
            """SELECT id, producto, fecha_programada, area_id,
                      COALESCE(estado,'')
                 FROM produccion_programada WHERE id=?""", (pid,)
        ).fetchone()
        if not pp:
            errores.append({'id': pid, 'error': 'producción no existe'})
            continue
        # No permitir reasignar producciones completadas o canceladas
        if pp[4].lower() in ('completado', 'cancelado'):
            errores.append({'id': pid, 'error': f'no se puede reasignar (estado: {pp[4]})'})
            continue

        prev_area = pp[3]
        # Conflicto: misma fecha + sala ya tomada → warning, no bloquea
        if new_area is not None:
            conf = c.execute("""
                SELECT id, producto FROM produccion_programada
                 WHERE fecha_programada=? AND area_id=? AND id<>?
                   AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
            """, (pp[2], new_area, pid)).fetchall()
            if conf:
                warnings.append({
                    'id': pid, 'producto': pp[1], 'fecha': pp[2],
                    'area_codigo': areas_map[new_area]['codigo'],
                    'choca_con': [{'id': x[0], 'producto': x[1]} for x in conf],
                })

        c.execute(
            "UPDATE produccion_programada SET area_id=? WHERE id=?",
            (new_area, pid)
        )
        # audit_log
        try:
            from audit_helpers import audit_log
            audit_log(
                c,
                accion='ASIGNAR_AREA_PROGRAMADA',
                tabla='produccion_programada',
                fila_id=pid,
                usuario=user,
                antes={'area_id': prev_area},
                despues={'area_id': new_area, 'producto': pp[1], 'fecha': pp[2]},
            )
        except Exception:
            pass
        ok_ids.append(pid)

    conn.commit()
    return jsonify({
        'ok': True,
        'asignados': len(ok_ids),
        'ids_ok': ok_ids,
        'errores': errores,
        'warnings': warnings,
    })


@bp.route('/api/planta/plan-semanal', methods=['GET'])
def planta_plan_semanal():
    """Vista consolidada del plan: cada producción ordenada por fecha,
    con días de inventario, consumo agregado y alertas.

    Sebastian (30-abr-2026): "siempre que calcula un producto suma los
    consumos de todas las programaciones previas a esa".

    Para cada producción programada:
      - dias_inventario_pt: días que duran las unidades PT con velocidad Shopify
      - mp_disponible_neto: stock_actual - sum(MP requerido por producciones PREVIAS)
      - alcanza_mp: bool (true si neto >= req actual)
      - mp_faltante: dict de materiales en déficit con cantidad
      - alerta_nivel: 'verde'|'amarillo'|'rojo' según días
      - presentaciones: lista de las del producto
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        dias = int(request.args.get('dias', 14))
    except (TypeError, ValueError):
        dias = 14
    conn = get_db(); c = conn.cursor()

    fecha_desde = datetime.now().strftime('%Y-%m-%d')
    fecha_hasta = (datetime.now() + timedelta(days=dias)).strftime('%Y-%m-%d')

    # 1) Producciones programadas en el rango
    producciones = c.execute("""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes, pp.estado,
               pp.area_id, ap.codigo as area_codigo, ap.nombre as area_nombre,
               fh.lote_size_kg, fh.imagen_url
        FROM produccion_programada pp
        LEFT JOIN areas_planta ap ON ap.id = pp.area_id
        LEFT JOIN formula_headers fh ON UPPER(TRIM(fh.producto_nombre)) = UPPER(TRIM(pp.producto))
        WHERE pp.fecha_programada BETWEEN ? AND ?
          AND COALESCE(pp.estado,'programado') NOT IN ('completado','cancelado')
        ORDER BY pp.fecha_programada ASC, pp.id ASC
    """, (fecha_desde, fecha_hasta)).fetchall()

    # 2) Consumo acumulado por material — fluye según orden temporal
    consumo_acumulado = {}  # material_id -> g acumulados
    items = []

    for row in producciones:
        prod_id, producto, fecha_prog, lotes, estado, area_id, area_codigo, area_nombre, lote_size, imagen_url = row
        lotes = lotes or 1
        mp_req = _calcular_mp_requerido(producto, lotes, c)

        # Para cada MP, calcular disponible NETO al momento de esta producción
        mp_status = []
        deficit = []
        for mat_id, req_g in mp_req.items():
            stock_total = _stock_mp(mat_id, c)
            ya_reservado = consumo_acumulado.get(mat_id, 0)
            disp_neto = stock_total - ya_reservado
            mat_nom_row = c.execute(
                "SELECT material_nombre FROM formula_items "
                "WHERE material_id=? LIMIT 1", (mat_id,)
            ).fetchone()
            mat_nom = (mat_nom_row[0] if mat_nom_row else mat_id)
            estado_mp = 'ok' if disp_neto >= req_g else ('justo' if disp_neto >= req_g * 0.5 else 'deficit')
            mp_status.append({
                'material_id': mat_id, 'material_nombre': mat_nom,
                'requerido_g': round(req_g),
                'stock_total_g': round(stock_total),
                'reservado_previo_g': round(ya_reservado),
                'disponible_neto_g': round(disp_neto),
                'estado': estado_mp,
            })
            if estado_mp == 'deficit':
                deficit.append({
                    'material_id': mat_id, 'material_nombre': mat_nom,
                    'falta_g': round(req_g - disp_neto),
                })
            # Acumular para siguientes
            consumo_acumulado[mat_id] = ya_reservado + req_g

        # Días de inventario PT
        # Buscar SKUs de este producto y velocidad de venta
        velocidad_dia = 0
        stock_pt = 0
        sku_rows = c.execute(
            "SELECT sku FROM sku_producto_map WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND COALESCE(activo,1)=1",
            (producto,)
        ).fetchall()
        for (sku,) in sku_rows:
            vel = c.execute("""
                SELECT COALESCE(SUM(cantidad),0)/60.0
                FROM ordenes_shopify_items WHERE sku=? AND fecha >= date('now', '-5 hours', '-60 days')
            """, (sku,)).fetchone() if False else None  # legacy, may not exist
            # Stock PT
            sp = c.execute(
                "SELECT COALESCE(SUM(unidades_disponible),0) FROM stock_pt WHERE sku=?",
                (sku,)
            ).fetchone()
            if sp:
                stock_pt += sp[0] or 0
        dias_inv = round(stock_pt / max(velocidad_dia, 0.01), 1) if velocidad_dia > 0 else None

        # Nivel de alerta por días de inventario
        if dias_inv is None:
            alerta_dias = 'gris'
        elif dias_inv < 10:
            alerta_dias = 'rojo'
        elif dias_inv < 20:
            alerta_dias = 'amarillo'
        else:
            alerta_dias = 'verde'

        # Presentaciones del producto
        pres_rows = c.execute("""
            SELECT id, etiqueta, volumen_ml, peso_g, envase_codigo
            FROM producto_presentaciones
            WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?)) AND activo=1
        """, (producto,)).fetchall()
        presentaciones = [{
            'id': p[0], 'etiqueta': p[1], 'volumen_ml': p[2],
            'peso_g': p[3], 'envase_codigo': p[4]
        } for p in pres_rows]

        # ¿Alcanzan TODAS las MP para esta producción?
        alcanza_mp = len(deficit) == 0

        items.append({
            'produccion_id': prod_id,
            'producto': producto,
            'imagen_url': imagen_url,
            'fecha_programada': fecha_prog,
            'lotes': lotes,
            'lote_size_kg': lote_size,
            'estado': estado,
            'area_codigo': area_codigo,
            'area_nombre': area_nombre,
            'stock_pt_unidades': stock_pt,
            'dias_inventario': dias_inv,
            'alerta_dias': alerta_dias,
            'alcanza_mp': alcanza_mp,
            'mp_status': mp_status,
            'mp_deficit': deficit,
            'presentaciones': presentaciones,
        })

    # KPIs globales
    total = len(items)
    n_rojo = sum(1 for it in items if it['alerta_dias'] == 'rojo')
    n_amar = sum(1 for it in items if it['alerta_dias'] == 'amarillo')
    n_sin_mp = sum(1 for it in items if not it['alcanza_mp'])
    return jsonify({
        'rango': {'desde': fecha_desde, 'hasta': fecha_hasta, 'dias': dias},
        'items': items,
        'kpis': {
            'total': total,
            'alerta_roja_dias': n_rojo,
            'alerta_amarilla_dias': n_amar,
            'sin_mp_suficiente': n_sin_mp,
        }
    })


@bp.route('/api/planta/aceptar-produccion/<int:produccion_id>', methods=['POST'])
def planta_aceptar_produccion(produccion_id):
    """Cascade automático cuando el operario "acepta" la producción.

    Sebastian (30-abr-2026): "lo selecciona le sale con la foto, y de una
    sale señalar envases, solicitar etiquetas, armado de goteros si
    requiere, aceptar producción se dispone para realizar, entonces
    automaticamente pasa a que el sistema decida en que area se hace y
    genere todo".

    Acciones:
      1. Si no tiene area_id → llamar a sugerir_area y asignar la mejor
      2. Crear tareas operativas: señalar envases, solicitar etiquetas,
         armar goteros (si presentación lo requiere)
      3. Programar envasado tentativo al día siguiente
      4. Notificar a Calidad que viene muestra micro pronto
      5. Devolver el plan completo

    Body opcional: {presentacion_id, lote_personalizado, fecha_envasado_override}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    conn = get_db(); c = conn.cursor()

    pp_row = c.execute("""
        SELECT id, producto, fecha_programada, lotes, area_id, estado
        FROM produccion_programada WHERE id=?
    """, (produccion_id,)).fetchone()
    if not pp_row:
        return jsonify({'error': 'Producción no existe'}), 404
    prod_id, producto, fecha_prog, lotes, area_id, estado = pp_row
    lotes = lotes or 1

    # Reemplazo MyBatch · fase 1 · BPM estricto: no se acepta producción sin un
    # MBR (receta maestra) aprobado. Se evalúa ANTES de cualquier mutación.
    # Solo aplica con EBR_MODE='strict' (default 'off' no bloquea nada).
    if EBR_MODE == 'strict':
        _mbr_ok = c.execute(
            "SELECT 1 FROM mbr_templates WHERE producto_nombre=? AND estado='aprobado' LIMIT 1",
            (producto,),
        ).fetchone()
        if not _mbr_ok:
            return jsonify({
                'error': f"No se puede aceptar: '{producto}' no tiene un MBR "
                         f"(receta maestra) aprobado. Creá y aprobá el MBR antes "
                         f"de fabricar (BPM/INVIMA).",
                'codigo': 'SIN_MBR_APROBADO',
            }), 409

    # (Batch C 3-jun) El candado multi-lote del Batch A se reemplazó por soporte
    # real: el hook de creación del EBR (abajo) genera N legajos cuando lotes>1.

    log = []
    # 1) Asignar área si no la tiene
    if not area_id:
        # Calcular lote_kg para sugerencia
        fh = c.execute(
            "SELECT lote_size_kg FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))",
            (producto,)
        ).fetchone()
        lote_kg = (fh[0] if fh else 0) * lotes if fh else 0
        if lote_kg <= 0:
            log.append('⚠ No se pudo sugerir área: producto sin lote_size_kg')
        else:
            # Reusar la lógica de sugerencia
            cap_min = lote_kg * 1.2
            tanques = c.execute("""
                SELECT codigo, area_codigo, capacidad_litros
                FROM equipos_planta
                WHERE activo=1 AND tipo IN ('tanque','marmita','olla')
                  AND capacidad_litros >= ?
                ORDER BY capacidad_litros ASC
            """, (cap_min,)).fetchall()
            por_area = {}
            for t in tanques:
                a = t[1]
                if a and (a not in por_area or t[2] < por_area[a]):
                    por_area[a] = t[2]
            # Tomar la primera area que pueda producir
            elegida = None
            for area_codigo, cap in por_area.items():
                area_chk = c.execute(
                    "SELECT id, nombre, puede_producir, activo FROM areas_planta WHERE codigo=?",
                    (area_codigo,)
                ).fetchone()
                if area_chk and area_chk[2] and area_chk[3]:
                    elegida = (area_chk[0], area_chk[1], area_codigo)
                    break
            if elegida:
                c.execute("UPDATE produccion_programada SET area_id=? WHERE id=?",
                          (elegida[0], produccion_id))
                area_id = elegida[0]
                log.append(f'✓ Área asignada: {elegida[1]} ({elegida[2]})')
            else:
                log.append('⚠ Sin tanque disponible — asignar manualmente')
    else:
        area_row = c.execute(
            "SELECT codigo, nombre FROM areas_planta WHERE id=?", (area_id,)
        ).fetchone()
        if area_row:
            log.append(f'✓ Área ya asignada: {area_row[1]}')

    # 2) Crear tareas operativas
    pres_id = d.get('presentacion_id')
    pres_etiqueta = ''
    requiere_gotero = False
    if pres_id:
        pr = c.execute(
            "SELECT etiqueta, envase_codigo, presentacion_codigo FROM producto_presentaciones WHERE id=?",
            (pres_id,)
        ).fetchone()
        if pr:
            pres_etiqueta = pr[0]
            # Heurística: si la presentación es suero/gotero o tiene "gotero" en etiqueta
            etl = (pr[0] or '').lower()
            cod = (pr[2] or '').lower()
            requiere_gotero = ('suero' in etl or 'gotero' in etl or cod.startswith('sue_') or cod.startswith('co_'))

    fecha_obj = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    tareas_creadas = []

    # Tarea 1: señalar envases
    cur = c.execute("""
        INSERT INTO tareas_operativas
          (titulo, descripcion, tipo, producto_relacionado, asignado_a,
           fecha_objetivo, estado, origen_tipo, origen_id, creado_por)
        VALUES (?, ?, 'recoleccion_mee', ?, ?, ?, 'pendiente', 'aceptar_produccion', ?, ?)
    """, (
        f'Señalar envases para {producto}' + (f' · {pres_etiqueta}' if pres_etiqueta else ''),
        f'Marcar y separar envases de bodega MEE para producción del {fecha_prog}.',
        producto, 'mayerlin,operarios', fecha_obj, produccion_id, user
    ))
    tareas_creadas.append({'id': cur.lastrowid, 'tipo': 'recoleccion_mee'})

    # Tarea 2: solicitar etiquetas
    cur = c.execute("""
        INSERT INTO tareas_operativas
          (titulo, descripcion, tipo, producto_relacionado, asignado_a,
           fecha_objetivo, estado, origen_tipo, origen_id, creado_por)
        VALUES (?, ?, 'recoleccion_mee', ?, ?, ?, 'pendiente', 'aceptar_produccion', ?, ?)
    """, (
        f'Solicitar etiquetas para {producto}',
        f'Confirmar etiquetas listas para envasado del {fecha_obj}. Si no están, alertar.',
        producto, 'camilo,catalina', fecha_obj, produccion_id, user
    ))
    tareas_creadas.append({'id': cur.lastrowid, 'tipo': 'etiquetas'})

    # Tarea 3: armar goteros (si aplica)
    if requiere_gotero:
        cur = c.execute("""
            INSERT INTO tareas_operativas
              (titulo, descripcion, tipo, producto_relacionado, asignado_a,
               fecha_objetivo, estado, origen_tipo, origen_id, creado_por)
            VALUES (?, ?, 'recoleccion_mee', ?, ?, ?, 'pendiente', 'aceptar_produccion', ?, ?)
        """, (
            f'Armar goteros para {producto}' + (f' · {pres_etiqueta}' if pres_etiqueta else ''),
            'Ensamblar goteros (cuerpo + bulbo) según presentación.',
            producto, 'camilo,operarios', fecha_obj, produccion_id, user
        ))
        tareas_creadas.append({'id': cur.lastrowid, 'tipo': 'armar_goteros'})

    log.append(f'✓ {len(tareas_creadas)} tarea(s) operativa(s) creada(s) para {fecha_obj}')

    # 3) Notificar a Calidad que viene muestra micro
    try:
        from blueprints.notif import push_notif_multi
        push_notif_multi(
            ['alejandro', 'sebastian'],
            'planta',
            f'🧪 Producción aceptada: {producto}',
            body=f'Lote programado para {fecha_prog} en área {area_id}. Muestra micro se solicitará al iniciar envasado.',
            link='/inventarios#programacion',
            remitente=user,
        )
        log.append('✓ Calidad notificada')
    except Exception as e:
        log.append(f'⚠ Notif no enviada: {e}')

    # 4) Marcar produccion como confirmada (estado='confirmada' opcional)
    c.execute(
        "UPDATE produccion_programada SET observaciones=COALESCE(observaciones,'')||'\n[ACEPTADA por '||?||' '||datetime('now', '-5 hours')||']' WHERE id=?",
        (user, produccion_id)
    )
    # Audit log · aceptar producción asigna área y crea tareas operativas
    # sobre produccion_programada. Trazabilidad obligatoria (quién aceptó/cuándo).
    try:
        audit_log(c, usuario=user, accion='ACEPTAR_PRODUCCION',
                  tabla='produccion_programada', registro_id=produccion_id,
                  despues={'producto': producto, 'area_id': area_id,
                           'tareas_creadas': len(tareas_creadas)})
    except Exception as _ae:
        logging.getLogger('programacion').warning('audit_log ACEPTAR_PRODUCCION fallo: %s', _ae)

    # Reemplazo MyBatch · fase 1 · crear/vincular el EBR (batch record) del lote
    # desde el MBR aprobado. EBR_MODE controla el comportamiento (off/warn/strict).
    # El bloqueo strict por falta de MBR ya se evaluó al inicio (antes de mutar).
    # Batch C · multi-lote: una producción con lotes>1 genera N legajos (1 BPR por
    # lote físico, exigencia BPM/INVIMA). lotes==1 mantiene el código 'PP<id>'
    # (compat); lotes>1 usa 'PP<id>-L1..N'. Cada uno idempotente por (prod_id,lote).
    ebr = None
    ebrs = []
    if EBR_MODE in ('warn', 'strict'):
        try:
            from blueprints.brd import crear_ebr_desde_mbr
            _n_lotes = lotes or 1
            for _i in range(1, _n_lotes + 1):
                _lote_ebr = f'PP{produccion_id}' if _n_lotes == 1 else f'PP{produccion_id}-L{_i}'
                _r = crear_ebr_desde_mbr(
                    c, producto_nombre=producto, lote=_lote_ebr,
                    produccion_id=produccion_id, usuario=user)
                ebrs.append(_r)
                if _r.get('ok'):
                    log.append('✓ EBR ' + str(_r.get('numero_op', '')) +
                               (' reusado' if _r.get('reusado') else
                                ' creado (' + str(_r.get('pasos', 0)) + ' pasos)'))
                    try:
                        audit_log(c, usuario=user, accion='CREAR_EBR_AUTO',
                                  tabla='ebr_ejecuciones', registro_id=_r.get('id'),
                                  despues={'produccion_id': produccion_id,
                                           'lote': _lote_ebr,
                                           'numero_op': _r.get('numero_op'),
                                           'reusado': bool(_r.get('reusado'))})
                    except Exception:
                        pass
                else:
                    log.append('⚠ EBR no creado: ' + str(_r.get('detail', _r.get('error'))))
            ebr = ebrs[0] if ebrs else None  # compat: respuesta 'ebr' = primero
        except Exception as _eebr:
            logging.getLogger('programacion').warning('crear EBR en aceptar fallo: %s', _eebr)

    conn.commit()

    return jsonify({
        'ok': True,
        'produccion_id': produccion_id,
        'producto': producto,
        'log': log,
        'tareas_creadas': tareas_creadas,
        'fecha_envasado_estimada': fecha_obj,
        'area_id': area_id,
        'ebr': ebr,
        'ebrs': ebrs,
    })


@bp.route('/api/planta/preflight/<int:produccion_id>/confirmar-limpieza', methods=['POST'])
def planta_preflight_confirmar_limpieza(produccion_id):
    """Operario confirma que acaba de hacer limpieza profunda en la sala
    asociada a esta producción. Registra evento + actualiza
    areas_planta.ultima_limpieza_profunda."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    nota = (d.get('nota') or '').strip() or 'Limpieza profunda confirmada vía preflight'
    conn = get_db(); c = conn.cursor()
    pp = c.execute(
        "SELECT area_id FROM produccion_programada WHERE id=?", (produccion_id,)
    ).fetchone()
    if not pp or not pp[0]:
        return jsonify({'error': 'Producción sin área asignada'}), 400
    area_id = pp[0]
    c.execute("UPDATE areas_planta SET ultima_limpieza_profunda=datetime('now', '-5 hours') WHERE id=?", (area_id,))
    c.execute("""
        INSERT INTO area_eventos (area_id, tipo, produccion_id, usuario, nota)
        VALUES (?, 'fin_limpieza', ?, ?, ?)
    """, (area_id, produccion_id, user, nota))
    conn.commit()
    return jsonify({'ok': True, 'mensaje': 'Limpieza profunda registrada'})
