# blueprints/financiero.py — extraído de index.py (Fase C)
import os
import json
import sqlite3
import hmac
import time
from datetime import datetime, date, timedelta
from flask import Blueprint, jsonify, request, Response, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from config import DB_PATH, COMPRAS_USERS, ADMIN_USERS, CONTADORA_USERS
from database import get_db, db_connect
from auth import _client_ip, _is_locked, _record_failure, _clear_attempts, _log_sec
from http_helpers import validate_money
from audit_helpers import audit_log
# ── "HOY" SIEMPRE EN HORA COLOMBIA (26-jul-2026) ─────────────────────────────────────────────
# Render corre en UTC y la empresa opera en Colombia (UTC-5). `_hoy_col()` del servidor salta al
# día siguiente después de las 7 de la tarde local, así que un pago registrado el 31 a las 19:30
# se guardaba con fecha del 1 y su `periodo` contable caía en el MES SIGUIENTE. Es dinero en el
# mes equivocado, y pasaba todas las noches de fin de mes.
# `api/tz_colombia.py` ya resolvía esto desde mayo y sólo lo usaba un módulo (M1: el helper
# canónico existe, hay que usarlo).
from tz_colombia import hoy_colombia as _hoy_col

from templates_py.rrhh_html import RRHH_HTML
from templates_py.compromisos_html import COMPROMISOS_HTML
from templates_py.home_html import HOME_HTML
from templates_py.hub_html import HUB_HTML
from templates_py.clientes_html import CLIENTES_HTML
from templates_py.calidad_html import CALIDAD_HTML
from templates_py.gerencia_html import GERENCIA_HTML
from templates_py.financiero_html import FINANCIERO_HTML
from templates_py.login_html import LOGIN_HTML
from templates_py.compras_html import COMPRAS_HTML
from templates_py.recepcion_html import RECEPCION_HTML
from templates_py.salida_html import SALIDA_HTML
from templates_py.solicitudes_html import SOLICITUDES_HTML
from templates_py.dashboard_html import DASHBOARD_HTML

bp = Blueprint('financiero', __name__)

@bp.route('/financiero')
def financiero_page():
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return redirect(url_for('core.login'))
    return Response(FINANCIERO_HTML, mimetype='text/html')

@bp.route('/api/financiero/ingresos', methods=['GET','POST'])
def handle_fin_ingresos():
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        if not d.get('concepto'):
            return jsonify({'error': 'Concepto requerido'}), 400
        # Money sanity validation · audit zero-error
        monto, err = validate_money(d.get('monto'), allow_zero=False, field_name='monto')
        if err:
            return jsonify(err), 400
        periodo = (d.get('fecha') or datetime.now().isoformat())[:7]
        usuario = session.get('compras_user','sistema')
        cur = c.execute("""INSERT INTO flujo_ingresos (fecha,empresa,concepto,categoria,monto,periodo,fuente,referencia,creado_por)
                     VALUES (?,?,?,?,?,?,?,?,?)""",
                  (d.get('fecha', datetime.now().isoformat()[:10]), d.get('empresa','HHA'),
                   d['concepto'], d.get('categoria','Ventas'), monto,
                   periodo, 'manual', d.get('referencia',''), usuario))
        audit_log(c, usuario=usuario, accion='CREAR_INGRESO_FIN',
                  tabla='flujo_ingresos', registro_id=cur.lastrowid,
                  despues={'concepto': d['concepto'][:80], 'monto': monto,
                           'empresa': d.get('empresa','HHA'),
                           'categoria': d.get('categoria','Ventas')},
                  detalle=f"+${monto:,.0f} {d.get('empresa','HHA')} · {d['concepto'][:60]}")
        conn.commit()
        return jsonify({'message': f"Ingreso de ${monto:,.0f} registrado"}), 201
    mes = request.args.get('mes')
    q = "SELECT id,fecha,empresa,concepto,categoria,monto,periodo,referencia FROM flujo_ingresos"
    params = []
    if mes: q += " WHERE periodo=?"; params.append(mes)
    q += " ORDER BY fecha DESC LIMIT 200"
    c.execute(q, params)
    cols = ['id','fecha','empresa','concepto','categoria','monto','periodo','referencia']
    return jsonify({'ingresos': [dict(zip(cols, r)) for r in c.fetchall()]})

@bp.route('/api/financiero/egresos', methods=['GET','POST'])
def handle_fin_egresos():
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        if not d.get('concepto'):
            return jsonify({'error': 'Concepto requerido'}), 400
        monto, err = validate_money(d.get('monto'), allow_zero=False, field_name='monto')
        if err:
            return jsonify(err), 400
        periodo = (d.get('fecha') or datetime.now().isoformat())[:7]
        usuario = session.get('compras_user','sistema')
        cur = c.execute("""INSERT INTO flujo_egresos (fecha,empresa,concepto,categoria,monto,periodo,fuente,referencia,creado_por)
                     VALUES (?,?,?,?,?,?,?,?,?)""",
                  (d.get('fecha', datetime.now().isoformat()[:10]), d.get('empresa','HHA'),
                   d['concepto'], d.get('categoria','MPs'), monto,
                   periodo, 'manual', d.get('referencia',''), usuario))
        audit_log(c, usuario=usuario, accion='CREAR_EGRESO_FIN',
                  tabla='flujo_egresos', registro_id=cur.lastrowid,
                  despues={'concepto': d['concepto'][:80], 'monto': monto,
                           'empresa': d.get('empresa','HHA'),
                           'categoria': d.get('categoria','MPs')},
                  detalle=f"-${monto:,.0f} {d.get('empresa','HHA')} · {d['concepto'][:60]}")
        conn.commit()
        return jsonify({'message': f"Egreso de ${monto:,.0f} registrado"}), 201
    mes = request.args.get('mes')
    q = "SELECT id,fecha,empresa,concepto,categoria,monto,periodo,referencia FROM flujo_egresos"
    params = []
    if mes: q += " WHERE periodo=?"; params.append(mes)
    q += " ORDER BY fecha DESC LIMIT 200"
    c.execute(q, params)
    cols = ['id','fecha','empresa','concepto','categoria','monto','periodo','referencia']
    return jsonify({'egresos': [dict(zip(cols, r)) for r in c.fetchall()]})

@bp.route('/api/financiero/kpis')
def financiero_kpis():
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # Auto-sync Shopify → flujo_ingresos antes de calcular KPIs.
    # Asi cuando Sebastian abre el dashboard, los pedidos Shopify recientes
    # ya cuentan como ingresos del mes (antes el card decia $0 mientras
    # Shopify mostraba $284.9M). Falla "best-effort" pero ahora con log
    # (Sebastián 25-may-2026 audit zero-error · antes fallaba silencioso
    # · si token Shopify vencía o rate limit, KPIs quedaban viejos sin
    # ninguna pista en logs).
    try:
        _sync_shopify_a_flujo_ingresos(conn, solo_pagados=True)
    except Exception as _e_sync:
        import logging as _lg_sync
        _lg_sync.getLogger('financiero').warning(
            'auto-sync shopify→flujo_ingresos fallo (KPIs pueden estar viejos): %s',
            _e_sync)
    # Auto-backfill de egresos mal categorizados como Espagiria que en
    # realidad son pagos a influencers (deben quedar como Animus). Idempotente.
    # Sebastian 2026-04-29: "todo esto creo son influencers deberia quedar
    # por animus no espagiria".
    try:
        _backfill_egresos_animus(conn)
    except Exception:
        pass
    # Colombia, no UTC (M24): al cerrar el mes, después de las 19:00 locales el tablero
    # saltaba al mes siguiente y mostraba los KPIs en cero con el mes todavía en curso.
    periodo_actual = _hoy_col().strftime('%Y-%m')
    # KPIs mes actual
    c.execute("SELECT COALESCE(SUM(monto),0), COUNT(*) FROM flujo_ingresos WHERE periodo=?", (periodo_actual,))
    ing_mes, ing_count = c.fetchone()
    c.execute("SELECT COALESCE(SUM(monto),0), COUNT(*) FROM flujo_egresos WHERE periodo=?", (periodo_actual,))
    egr_mes, egr_count = c.fetchone()
    # ⚠ `gerencia_inputs.saldo_caja` es el efectivo de la EMPRESA y lo TECLEA gerencia una vez
    # al mes. Se mostraba pelado, al lado de tres números calculados en vivo, tomando el último
    # período que exista -- así que un número de hace seis semanas se leía igual de fresco que
    # el ingreso de hoy. El dato no cambia; lo que cambia es que ahora dice DE CUÁNDO es y si
    # está vencido, porque un indicador que alguien tiene que acordarse de actualizar termina
    # viejo y nadie lo nota (M109/M124).
    c.execute("SELECT saldo_caja, periodo FROM gerencia_inputs ORDER BY periodo DESC LIMIT 1")
    row = c.fetchone()
    saldo_caja = row[0] if row else 0
    saldo_caja_periodo = (row[1] if row else '') or ''
    saldo_caja_vigente = bool(saldo_caja_periodo == periodo_actual)
    # La CAJA MENOR es otra cosa y sí es verificable: es lo que se cuenta en el arqueo. Sale del
    # helper canónico, el MISMO contra el que se autorizan los pagos (M1/M148) -- nunca de un
    # SUM propio, que volvería a divergir.
    try:
        from blueprints.animus import caja_saldo as _caja_saldo
        caja_menor = float(_caja_saldo(conn) or 0)
        caja_menor_ok = True
    except Exception as _e_cm:
        # Lo que no se pudo medir se declara: un cero inventado se lee como "no hay plata" y
        # significa lo contrario, "no se miró" (M100).
        # ⚠ Este módulo NO tiene un `log` global: escribir `log.warning` acá habría lanzado un
        # NameError DENTRO del except y tumbado el endpoint entero -- el fallback que existe
        # para no caerse, tirando la página abajo. Se usa el idiom del propio archivo.
        import logging as _lg_cm
        _lg_cm.getLogger('financiero').warning(
            'no pude leer la caja menor para los KPI: %s', _e_cm)
        caja_menor = None
        caja_menor_ok = False
    # Desglose por categoría mes actual
    c.execute("SELECT categoria, SUM(monto) as total FROM flujo_ingresos WHERE periodo=? GROUP BY categoria ORDER BY total DESC", (periodo_actual,))
    desglose_ing = [{'categoria': r[0], 'total': r[1]} for r in c.fetchall()]
    c.execute("SELECT categoria, SUM(monto) as total FROM flujo_egresos WHERE periodo=? GROUP BY categoria ORDER BY total DESC", (periodo_actual,))
    desglose_egr = [{'categoria': r[0], 'total': r[1]} for r in c.fetchall()]
    # Histórico 6 meses · audit zero-error 2-may-2026
    # Antes: 12 queries (2 × 6 meses) en loop. Ahora: 2 queries pre-agregadas.
    from datetime import date as _d
    hoy_d = _d.today()
    periodos = []
    for i in range(5, -1, -1):
        mes = hoy_d.month - i
        anio = hoy_d.year
        while mes <= 0: mes += 12; anio -= 1
        periodos.append(f"{anio}-{mes:02d}")
    placeholders = ','.join(['?'] * len(periodos))
    ing_map = dict(c.execute(
        f"SELECT periodo, COALESCE(SUM(monto),0) FROM flujo_ingresos "
        f"WHERE periodo IN ({placeholders}) GROUP BY periodo",
        periodos
    ).fetchall())
    egr_map = dict(c.execute(
        f"SELECT periodo, COALESCE(SUM(monto),0) FROM flujo_egresos "
        f"WHERE periodo IN ({placeholders}) GROUP BY periodo",
        periodos
    ).fetchall())
    historico = [
        {'periodo': p, 'ingresos': ing_map.get(p, 0), 'egresos': egr_map.get(p, 0)}
        for p in periodos
    ]
    # Shopify real-time (DTC directo) — no está en flujo_ingresos manual
    try:
        c.execute("SELECT COALESCE(SUM(total),0), COUNT(*) FROM animus_shopify_orders WHERE creado_en LIKE ?",
                  (periodo_actual+'%',))
        shp_row = c.fetchone()
        shopify_mes = round(shp_row[0] or 0, 0)
        shopify_pedidos = shp_row[1] or 0
        # YTD Shopify
        anio_ini = datetime.now().strftime('%Y')
        c.execute("SELECT COALESCE(SUM(total),0) FROM animus_shopify_orders WHERE creado_en LIKE ?",
                  (anio_ini+'%',))
        shopify_anio = round(c.fetchone()[0] or 0, 0)
    except Exception:
        shopify_mes = 0; shopify_pedidos = 0; shopify_anio = 0
    return jsonify({'ing_mes': ing_mes, 'ing_count': ing_count, 'egr_mes': egr_mes, 'egr_count': egr_count,
                    'saldo_caja': saldo_caja,
                    'saldo_caja_periodo': saldo_caja_periodo,
                    'saldo_caja_vigente': saldo_caja_vigente,
                    'caja_menor': caja_menor, 'caja_menor_ok': caja_menor_ok,
                    'desglose_ing': desglose_ing, 'desglose_egr': desglose_egr,
                    'historico': historico, 'periodo': periodo_actual,
                    'shopify_mes': shopify_mes, 'shopify_pedidos': shopify_pedidos, 'shopify_anio': shopify_anio})

@bp.route('/api/financiero/flujo-mensual')
def financiero_flujo_mensual():
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT periodo, SUM(monto) FROM flujo_ingresos GROUP BY periodo ORDER BY periodo")
    ings = {r[0]: r[1] for r in c.fetchall()}
    c.execute("SELECT periodo, SUM(monto) FROM flujo_egresos GROUP BY periodo ORDER BY periodo")
    egrs = {r[0]: r[1] for r in c.fetchall()}
    periodos = sorted(set(list(ings.keys()) + list(egrs.keys())))
    meses = [{'periodo': p, 'ingresos': ings.get(p, 0), 'egresos': egrs.get(p, 0)} for p in periodos]
    return jsonify({'meses': meses})


@bp.route('/api/financiero/mom-12-meses')
def financiero_mom_12_meses():
    """Retorna últimos 12 meses con ingresos/egresos/margen/MoM%.

    Diseñado para chart de tendencia + tabla de variación. Genera SIEMPRE 12
    períodos contiguos hacia atrás desde el mes actual (incluso si están
    vacíos · permite ver la curva sin huecos).

    Stats extra incluídas para análisis rápido:
    - mejor_mes / peor_mes (por margen)
    - margen_promedio últimos 12 meses
    - mejor/peor categoría de egresos del último mes
    """
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    from datetime import datetime as _dt
    conn = get_db(); c = conn.cursor()

    # Generar lista de 12 períodos desde mes actual hacia atrás
    hoy = _dt.now()
    periodos = []
    for i in range(11, -1, -1):  # de hace 11 meses hasta el mes actual
        anio = hoy.year
        mes = hoy.month - i
        while mes <= 0:
            mes += 12; anio -= 1
        periodos.append(f"{anio:04d}-{mes:02d}")

    # Bulk fetch de los 12 períodos
    placeholders = ','.join('?' for _ in periodos)
    try:
        ings_rows = c.execute(
            f"SELECT periodo, COALESCE(SUM(monto),0) FROM flujo_ingresos "
            f"WHERE periodo IN ({placeholders}) GROUP BY periodo",
            periodos
        ).fetchall()
        ings = {r[0]: float(r[1] or 0) for r in ings_rows}
    except Exception:
        ings = {}
    try:
        egrs_rows = c.execute(
            f"SELECT periodo, COALESCE(SUM(monto),0) FROM flujo_egresos "
            f"WHERE periodo IN ({placeholders}) GROUP BY periodo",
            periodos
        ).fetchall()
        egrs = {r[0]: float(r[1] or 0) for r in egrs_rows}
    except Exception:
        egrs = {}

    # Construir lista contigua con margen y MoM
    meses = []
    prev_ing = None
    for p in periodos:
        ing = ings.get(p, 0.0)
        egr = egrs.get(p, 0.0)
        margen = ing - egr
        margen_pct = (margen / ing * 100) if ing > 0 else 0
        mom_pct = None
        if prev_ing is not None and prev_ing > 0:
            mom_pct = (ing - prev_ing) / prev_ing * 100
        meses.append({
            'periodo': p,
            'ingresos': ing,
            'egresos': egr,
            'margen': margen,
            'margen_pct': round(margen_pct, 1),
            'mom_pct': round(mom_pct, 1) if mom_pct is not None else None,
        })
        prev_ing = ing

    # Stats
    meses_con_data = [m for m in meses if m['ingresos'] > 0 or m['egresos'] > 0]
    stats = {}
    if meses_con_data:
        sorted_by_margen = sorted(meses_con_data, key=lambda x: x['margen'], reverse=True)
        stats['mejor_mes'] = sorted_by_margen[0]['periodo']
        stats['mejor_mes_margen'] = sorted_by_margen[0]['margen']
        stats['peor_mes'] = sorted_by_margen[-1]['periodo']
        stats['peor_mes_margen'] = sorted_by_margen[-1]['margen']
        margenes = [m['margen'] for m in meses_con_data]
        stats['margen_promedio'] = sum(margenes) / len(margenes)
        # Top categoría de egresos último mes (último período del rango)
        ultimo_periodo = periodos[-1]
        try:
            top_cat = c.execute("""
                SELECT COALESCE(categoria,'Otro'), SUM(monto)
                FROM flujo_egresos WHERE periodo=?
                GROUP BY categoria ORDER BY 2 DESC LIMIT 1
            """, (ultimo_periodo,)).fetchone()
            if top_cat and top_cat[1]:
                stats['top_categoria_egreso'] = top_cat[0]
                stats['top_categoria_egreso_monto'] = float(top_cat[1])
        except Exception:
            pass

    return jsonify({
        'meses': meses,
        'stats': stats,
        'rango': {'desde': periodos[0], 'hasta': periodos[-1]},
    })


@bp.route('/api/financiero/mes-detalle')
def financiero_mes_detalle():
    """Drill-down: detalle de un mes por categoría con top items.

    Query params:
      periodo · YYYY-MM (requerido)
      tipo    · 'ingresos' | 'egresos' (default 'egresos')

    Devuelve:
      total · suma del mes
      categorias · [{categoria, monto, count, pct, top_items: [{fecha,concepto,monto,referencia,fuente}]}]
    """
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    periodo = (request.args.get('periodo') or '').strip()
    tipo = (request.args.get('tipo') or 'egresos').strip().lower()
    if not periodo or len(periodo) != 7 or '-' not in periodo:
        return jsonify({'error': 'periodo requerido (YYYY-MM)'}), 400
    if tipo not in ('ingresos', 'egresos'):
        return jsonify({'error': "tipo debe ser 'ingresos' o 'egresos'"}), 400

    tabla = 'flujo_ingresos' if tipo == 'ingresos' else 'flujo_egresos'
    conn = get_db(); c = conn.cursor()
    # Total del mes
    try:
        total = c.execute(
            f"SELECT COALESCE(SUM(monto),0), COUNT(*) FROM {tabla} WHERE periodo=?",
            (periodo,)
        ).fetchone()
        total_monto = float(total[0] or 0)
        total_count = int(total[1] or 0)
    except Exception as e:
        return jsonify({'error': f'fallo lectura {tabla}: {str(e)[:200]}'}), 500
    # Agrupar por categoría
    try:
        cat_rows = c.execute(
            f"""SELECT COALESCE(NULLIF(TRIM(categoria),''),'Sin categoría') as cat,
                       COALESCE(SUM(monto),0) as total, COUNT(*) as n
                FROM {tabla} WHERE periodo=?
                GROUP BY cat ORDER BY total DESC""",
            (periodo,)
        ).fetchall()
    except Exception:
        cat_rows = []
    categorias = []
    for cat, monto_cat, n_cat in cat_rows:
        pct = (float(monto_cat or 0) / total_monto * 100) if total_monto > 0 else 0
        # Top 5 items de esta categoría
        try:
            items = c.execute(
                f"""SELECT fecha, concepto, monto, COALESCE(referencia,''),
                            COALESCE(fuente,'manual')
                    FROM {tabla}
                    WHERE periodo=? AND
                          COALESCE(NULLIF(TRIM(categoria),''),'Sin categoría') = ?
                    ORDER BY monto DESC LIMIT 5""",
                (periodo, cat)
            ).fetchall()
            top_items = [{
                'fecha': r[0], 'concepto': (r[1] or '')[:120],
                'monto': float(r[2] or 0),
                'referencia': r[3], 'fuente': r[4],
            } for r in items]
        except Exception:
            top_items = []
        categorias.append({
            'categoria': cat, 'monto': float(monto_cat or 0),
            'count': int(n_cat or 0), 'pct': round(pct, 1),
            'top_items': top_items,
        })
    return jsonify({
        'periodo': periodo, 'tipo': tipo,
        'total': total_monto, 'count': total_count,
        'categorias': categorias,
    })


@bp.route('/api/financiero/categoria-trend')
def financiero_categoria_trend():
    """Tendencia de UNA categoría a lo largo de 12 meses.

    Query params:
      categoria · nombre exacto (ej. 'MPs') · requerido
      tipo      · 'ingresos' | 'egresos' (default 'egresos')

    Devuelve:
      meses · 12 períodos contiguos con {periodo, monto, count}
      stats · {promedio, max_mes, min_mes, total_12m, ultimo_vs_promedio_pct}
    """
    u = session.get('compras_user','')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    categoria = (request.args.get('categoria') or '').strip()
    tipo = (request.args.get('tipo') or 'egresos').strip().lower()
    if not categoria:
        return jsonify({'error': 'categoria requerida'}), 400
    if tipo not in ('ingresos', 'egresos'):
        return jsonify({'error': "tipo debe ser 'ingresos' o 'egresos'"}), 400
    tabla = 'flujo_ingresos' if tipo == 'ingresos' else 'flujo_egresos'

    from datetime import datetime as _dt
    conn = get_db(); c = conn.cursor()
    # 12 períodos hacia atrás desde el mes actual
    hoy = _dt.now()
    periodos = []
    for i in range(11, -1, -1):
        anio = hoy.year
        mes = hoy.month - i
        while mes <= 0:
            mes += 12; anio -= 1
        periodos.append(f"{anio:04d}-{mes:02d}")

    # Match flexible: si categoria es 'Sin categoría' busca NULL/empty
    if categoria == 'Sin categoría':
        where_cat = "(categoria IS NULL OR TRIM(categoria) = '')"
        params_cat = []
    else:
        where_cat = "categoria = ?"
        params_cat = [categoria]

    placeholders = ','.join('?' for _ in periodos)
    try:
        rows = c.execute(
            f"""SELECT periodo, COALESCE(SUM(monto),0), COUNT(*)
                FROM {tabla}
                WHERE periodo IN ({placeholders}) AND {where_cat}
                GROUP BY periodo""",
            periodos + params_cat
        ).fetchall()
    except Exception:
        rows = []
    by_periodo = {r[0]: (float(r[1] or 0), int(r[2] or 0)) for r in rows}

    meses = []
    for p in periodos:
        monto, n = by_periodo.get(p, (0.0, 0))
        meses.append({'periodo': p, 'monto': monto, 'count': n})

    # Stats
    montos = [m['monto'] for m in meses]
    montos_no_zero = [x for x in montos if x > 0]
    stats = {
        'total_12m': sum(montos),
        'promedio': sum(montos_no_zero) / len(montos_no_zero) if montos_no_zero else 0,
    }
    if montos_no_zero:
        max_idx = montos.index(max(montos_no_zero))
        min_idx = montos.index(min(montos_no_zero))
        stats['max_mes'] = meses[max_idx]['periodo']
        stats['max_monto'] = meses[max_idx]['monto']
        stats['min_mes'] = meses[min_idx]['periodo']
        stats['min_monto'] = meses[min_idx]['monto']
        # Último mes vs promedio (cuanto se desvió este mes del promedio)
        ultimo = montos[-1]
        if stats['promedio'] > 0:
            stats['ultimo_vs_promedio_pct'] = round((ultimo - stats['promedio']) / stats['promedio'] * 100, 1)
        else:
            stats['ultimo_vs_promedio_pct'] = None
    return jsonify({
        'categoria': categoria, 'tipo': tipo,
        'meses': meses, 'stats': stats,
    })

@bp.route('/api/financiero/config', methods=['GET','POST'])
def financiero_config():
    if 'compras_user' not in session or session.get('compras_user','') not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        usuario = session.get('compras_user','sistema')
        # Capturar antes para audit
        antes = {}
        for clave in d.keys():
            row = c.execute("SELECT valor FROM flujo_config WHERE clave=?", (clave,)).fetchone()
            antes[clave] = row[0] if row else None
        for clave, valor in d.items():
            c.execute("INSERT INTO flujo_config (clave,valor,descripcion) VALUES (?,?,?) ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor", (clave, str(valor), ''))
        audit_log(c, usuario=usuario, accion='ACTUALIZAR_CONFIG_FIN',
                  tabla='flujo_config', registro_id=None,
                  antes=antes, despues={k: str(v) for k,v in d.items()},
                  detalle=f"{len(d)} parametros actualizados: {', '.join(list(d.keys())[:5])}")
        conn.commit()
        return jsonify({'message': f'{len(d)} parámetros actualizados'})
    c.execute("SELECT clave, valor FROM flujo_config ORDER BY clave")
    config = {r[0]: r[1] for r in c.fetchall()}
    return jsonify({'config': config})

@bp.route('/api/financiero/importar-ocs', methods=['POST'])
def financiero_importar_ocs():
    if 'compras_user' not in session or session.get('compras_user','') not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # Traer OCs recibidas que no estén ya importadas
    # FIX 10-jun audit · (a) GROUP BY incompleto → 500 en PG (numero_oc es UNIQUE, no PK)
    # · agregar todas las columnas no agregadas. (b) IVA: el egreso debe ser el valor_total
    # de la OC (ya incluye IVA · compras.py *1.19), no la suma de subtotales SIN IVA →
    # priorizar valor_total con NULLIF; SUM queda como fallback si la OC no tiene total.
    c.execute("""SELECT oc.numero_oc, oc.fecha, oc.proveedor,
                        COALESCE(NULLIF(oc.valor_total,0), SUM(i.cantidad_g * i.precio_unitario), 0) as total
                 FROM ordenes_compra oc
                 LEFT JOIN ordenes_compra_items i ON oc.numero_oc=i.numero_oc
                 WHERE oc.estado='Recibida'
                 AND oc.numero_oc NOT IN (SELECT referencia FROM flujo_egresos WHERE referencia LIKE 'OC-%')
                 GROUP BY oc.numero_oc, oc.fecha, oc.proveedor, oc.valor_total""")
    ocs = c.fetchall()
    usuario = session.get('compras_user','sistema')
    importadas = 0
    total_importado = 0.0
    for numero_oc, fecha, proveedor, total in ocs:
        if total and total > 0:
            periodo = (fecha or datetime.now().isoformat())[:7]
            c.execute("""INSERT INTO flujo_egresos (fecha,empresa,concepto,categoria,monto,periodo,fuente,referencia,creado_por)
                         VALUES (?,?,?,?,?,?,?,?,?)""",
                      (fecha[:10] if fecha else datetime.now().isoformat()[:10],
                       'ESPAGIRIA', f'OC {numero_oc} — {proveedor or ""}',
                       'MPs', float(total), periodo, 'automatico', numero_oc, usuario))
            importadas += 1
            total_importado += float(total)
    if importadas:
        audit_log(c, usuario=usuario, accion='IMPORTAR_OCS_FIN',
                  tabla='flujo_egresos', registro_id=None,
                  despues={'ocs_importadas': importadas, 'total_egreso': total_importado},
                  detalle=f"Importacion masiva · {importadas} OCs · ${total_importado:,.0f}")
    conn.commit()
    return jsonify({'message': f'{importadas} OC(s) importadas como egresos'})

@bp.route('/api/financiero/conciliacion-bancaria/preview', methods=['POST'])
def conciliacion_preview():
    """Conciliacion bancaria: recibe lineas del extracto bancario,
    intenta hacer match contra flujo_ingresos/egresos por monto+fecha.

    Body: {extracto_lineas: [{fecha, monto, descripcion, tipo: 'C'|'D'}]}

    Devuelve: cada linea con su match sugerido (si lo encuentra) +
    confidence score. NO escribe nada — es preview.
    """
    if 'compras_user' not in session or session.get('compras_user','') not in (ADMIN_USERS | CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    d = request.get_json() or {}
    lineas = d.get('extracto_lineas', [])
    conn = get_db(); c = conn.cursor()
    resultados = []
    for L in lineas:
        fecha = L.get('fecha', '')[:10]
        monto = abs(float(L.get('monto', 0)))
        tipo = (L.get('tipo','') or '').upper()
        desc = (L.get('descripcion','') or '').strip()
        if not fecha or monto == 0:
            resultados.append({'linea': L, 'match': None, 'razon': 'falta fecha o monto'})
            continue
        # Buscar en ingresos/egresos cercanos en fecha (+/- 3 dias) y monto exacto
        tabla = 'flujo_ingresos' if tipo == 'C' else 'flujo_egresos'
        match = c.execute(f"""
            SELECT id, fecha, concepto, monto, referencia, fuente,
                   ABS(julianday(fecha) - julianday(?)) as dias_diff
            FROM {tabla}
            WHERE ABS(monto - ?) < 0.01
              AND ABS(julianday(fecha) - julianday(?)) <= 3
            ORDER BY dias_diff ASC LIMIT 1
        """, (fecha, monto, fecha)).fetchone()
        if match:
            cols = ['id','fecha','concepto','monto','referencia','fuente','dias_diff']
            resultados.append({
                'linea': L,
                'match': dict(zip(cols, match)),
                'tabla': tabla,
                'confidence': 'alta' if match[6] == 0 else ('media' if match[6] <= 1 else 'baja'),
            })
        else:
            resultados.append({'linea': L, 'match': None,
                               'razon': 'sin coincidencia en fecha+monto'})
    return jsonify({'resultados': resultados,
                    'lineas_total': len(lineas),
                    'matched': sum(1 for r in resultados if r['match'])})


@bp.route('/api/financiero/pnl-por-empresa')
def pnl_por_empresa():
    """P&L separado por empresa del holding (HHA / Espagiria / Animus).

    Antes el P&L principal sumaba todo. Ahora desglosado para que Sebastian
    vea la salud de cada negocio del holding individualmente. #17 cerrado.
    """
    if 'compras_user' not in session or session.get('compras_user','') not in (ADMIN_USERS | CONTADORA_USERS):
        return jsonify({'error': 'No autorizado'}), 401
    desde = request.args.get('desde', _hoy_col().replace(day=1).isoformat())
    hasta = request.args.get('hasta', _hoy_col().isoformat())

    conn = get_db(); c = conn.cursor()
    empresas = ['HHA', 'ESPAGIRIA', 'ANIMUS']
    out = {}
    total_ingresos = total_egresos = 0.0
    for emp in empresas:
        try:
            ing = c.execute("""SELECT COALESCE(SUM(monto),0) FROM flujo_ingresos
                              WHERE UPPER(TRIM(empresa))=? AND fecha BETWEEN ? AND ?""",
                            (emp, desde, hasta)).fetchone()[0] or 0
            egr = c.execute("""SELECT COALESCE(SUM(monto),0) FROM flujo_egresos
                              WHERE UPPER(TRIM(empresa))=? AND fecha BETWEEN ? AND ?""",
                            (emp, desde, hasta)).fetchone()[0] or 0

            # Desglose ingresos por categoria
            ing_breakdown = {}
            for row in c.execute("""SELECT categoria, COALESCE(SUM(monto),0)
                                    FROM flujo_ingresos
                                    WHERE UPPER(TRIM(empresa))=? AND fecha BETWEEN ? AND ?
                                    GROUP BY categoria""", (emp, desde, hasta)).fetchall():
                ing_breakdown[row[0] or 'Sin categoria'] = row[1]

            # Desglose egresos por categoria
            egr_breakdown = {}
            for row in c.execute("""SELECT categoria, COALESCE(SUM(monto),0)
                                    FROM flujo_egresos
                                    WHERE UPPER(TRIM(empresa))=? AND fecha BETWEEN ? AND ?
                                    GROUP BY categoria""", (emp, desde, hasta)).fetchall():
                egr_breakdown[row[0] or 'Sin categoria'] = row[1]

            ebitda = ing - egr
            margen = round((ebitda / ing * 100), 1) if ing > 0 else None
            out[emp] = {
                'ingresos': ing,
                'egresos': egr,
                'ebitda': ebitda,
                'margen_pct': margen,
                'ingresos_por_categoria': ing_breakdown,
                'egresos_por_categoria': egr_breakdown,
            }
            total_ingresos += ing
            total_egresos += egr
        except Exception:
            out[emp] = {'ingresos': 0, 'egresos': 0, 'ebitda': 0, 'margen_pct': None}

    out['TOTAL_HOLDING'] = {
        'ingresos': total_ingresos,
        'egresos': total_egresos,
        'ebitda': total_ingresos - total_egresos,
        'margen_pct': round(((total_ingresos - total_egresos) / total_ingresos * 100), 1)
                      if total_ingresos > 0 else None,
    }
    out['_periodo'] = {'desde': desde, 'hasta': hasta}
    return jsonify(out)


def _backfill_egresos_animus(conn):
    """Recategoriza egresos viejos mal etiquetados como Espagiria que en
    realidad son pagos a influencers (deben quedar como Animus).

    Detecta via dos señales fuertes (no usa keywords ambiguas):
      1. La referencia (numero_oc) tiene una fila en pagos_influencers.
      2. La solicitud de compra origen tiene influencer_id NOT NULL.

    Idempotente: solo actualiza filas que SIGUEN como 'Espagiria'.
    Falla silenciosa si las tablas legacy no existen.

    Returns: count de filas actualizadas.
    """
    c = conn.cursor()
    try:
        cur1 = c.execute("""
            UPDATE flujo_egresos
            SET empresa='Animus'
            WHERE LOWER(COALESCE(empresa,''))='espagiria'
              AND referencia IS NOT NULL
              AND referencia IN (
                SELECT DISTINCT numero_oc FROM pagos_influencers
                WHERE numero_oc IS NOT NULL AND numero_oc != ''
              )
        """)
        n1 = cur1.rowcount or 0
    except sqlite3.OperationalError:
        n1 = 0  # tabla pagos_influencers no existe en deploy viejo
    try:
        cur2 = c.execute("""
            UPDATE flujo_egresos
            SET empresa='Animus'
            WHERE LOWER(COALESCE(empresa,''))='espagiria'
              AND referencia IS NOT NULL
              AND referencia IN (
                SELECT DISTINCT numero_oc FROM solicitudes_compra
                WHERE influencer_id IS NOT NULL AND numero_oc IS NOT NULL
              )
        """)
        n2 = cur2.rowcount or 0
    except sqlite3.OperationalError:
        n2 = 0  # columna influencer_id no existe
    if n1 or n2:
        conn.commit()
    return n1 + n2


def _sync_shopify_a_flujo_ingresos(conn, solo_pagados=True, desde_fecha='', dry_run=False):
    """Helper compartido: sincroniza pedidos Shopify a flujo_ingresos.

    Idempotente: ordenes con flujo_synced=1 se saltan; si ya existe un
    ingreso con la misma referencia (SHOPIFY-<id>), tambien se salta y
    solo se actualiza el link.

    Usa indices numericos en lugar de Row factory para no exigir que el
    caller configure conn.row_factory — asi se puede llamar desde
    financiero_kpis sin contaminar las queries siguientes que esperan
    tuples.

    Returns dict con: ok, pendientes, importadas, total_importado, error.
    """
    c = conn.cursor()
    # Verificar que la columna flujo_synced existe (migracion 37)
    try:
        c.execute("SELECT flujo_synced FROM animus_shopify_orders LIMIT 1")
    except sqlite3.OperationalError:
        return {'ok': False, 'error': 'Migracion #37 no aplicada (flujo_synced ausente)',
                'pendientes': 0, 'importadas': 0, 'total_importado': 0}

    where = ['(flujo_synced=0 OR flujo_synced IS NULL)']
    params = []
    if solo_pagados:
        where.append("LOWER(COALESCE(estado_pago,'')) IN ('paid','pagado','complete','captured','partially_paid')")
    if desde_fecha:
        where.append("creado_en >= ?")
        params.append(desde_fecha)

    # SELECT con orden FIJO de columnas: id(0), shopify_id(1), nombre(2),
    # total(3), moneda(4), creado_en(5), ciudad(6)
    sql = f"""SELECT id, shopify_id, nombre, total, moneda, creado_en, ciudad
              FROM animus_shopify_orders
              WHERE {' AND '.join(where)}
              ORDER BY creado_en"""
    pendientes = c.execute(sql, params).fetchall()
    if not pendientes:
        return {'ok': True, 'pendientes': 0, 'importadas': 0, 'total_importado': 0}

    if dry_run:
        return {'ok': True, 'dry_run': True, 'pendientes': len(pendientes),
                'importadas': 0,
                'total_importado': sum(float(r[3] or 0) for r in pendientes)}

    importadas = 0
    total_importado = 0.0
    for row in pendientes:
        order_id = row[0]
        shopify_id = row[1] or ''
        total = float(row[3] or 0)
        if total <= 0:
            continue
        # FIX 10-jun audit · RACE multi-worker (3 Gunicorn + trigger on-demand desde KPIs):
        # reclamar el pedido ATÓMICAMENTE antes de insertar. El CAS marca flujo_synced=1
        # solo si seguía en 0 → un único worker gana cada pedido (rowcount==1); el resto
        # ve rowcount==0 y salta. Evita el doble INSERT que inflaba los ingresos (la tabla
        # flujo_ingresos no tiene UNIQUE en referencia, así que el check-then-insert no bastaba).
        if c.execute(
            "UPDATE animus_shopify_orders SET flujo_synced=1 "
            "WHERE id=? AND (flujo_synced=0 OR flujo_synced IS NULL)",
            (order_id,)
        ).rowcount != 1:
            continue  # otro worker ya está sincronizando este pedido
        fecha = (row[5] or datetime.now().isoformat())[:10]
        periodo = fecha[:7]
        nombre = row[2] or ''
        concepto = f'Shopify {shopify_id}' + (f' — {nombre[:40]}' if nombre else '')
        ref = f'SHOPIFY-{shopify_id}' if shopify_id else f'SHOPIFY-{order_id}'
        existente = c.execute(
            "SELECT id FROM flujo_ingresos WHERE referencia=?", (ref,)
        ).fetchone()
        if existente:
            c.execute(
                "UPDATE animus_shopify_orders SET flujo_ingreso_id=? WHERE id=?",
                (existente[0], order_id)
            )
            continue
        c.execute("""INSERT INTO flujo_ingresos
                     (fecha, empresa, concepto, categoria, monto, periodo,
                      fuente, referencia, creado_por)
                     VALUES (?,?,?,?,?,?,?,?,?)""",
                  (fecha, 'ANIMUS', concepto, 'Ventas Shopify',
                   total, periodo, 'shopify_auto', ref, 'sistema_sync'))
        ing_id = c.lastrowid
        c.execute(
            "UPDATE animus_shopify_orders SET flujo_ingreso_id=? WHERE id=?",
            (ing_id, order_id)
        )
        importadas += 1
        total_importado += total
    conn.commit()
    return {'ok': True, 'pendientes': len(pendientes), 'importadas': importadas,
            'total_importado': total_importado}


def _start_shopify_ingresos_background_loop():
    """Thread daemon que cada SHOPIFY_INGRESOS_SYNC_INTERVAL_MIN min sincroniza
    pedidos Shopify nuevos a flujo_ingresos sin que nadie tenga la pantalla
    abierta. Default 30 min. <=0 desactiva.

    Resuelve la queja Sebastian (29-abr-2026): "lo de shopy no lo esta
    registrando como ingresos" — antes solo se sincronizaba si Sebastian
    presionaba un boton manual o llamaba el endpoint.
    """
    import threading, os, time as _t
    if getattr(_start_shopify_ingresos_background_loop, '_running', False):
        return
    try:
        interval_min = int(os.environ.get('SHOPIFY_INGRESOS_SYNC_INTERVAL_MIN', '30'))
    except ValueError:
        interval_min = 30
    if interval_min <= 0:
        return
    _start_shopify_ingresos_background_loop._running = True

    def _worker():
        from config import DB_PATH
        import sqlite3 as _sql
        _t.sleep(45)  # delay inicial
        while True:
            try:
                local_conn = db_connect(timeout=30)
                try:
                    _sync_shopify_a_flujo_ingresos(local_conn, solo_pagados=True)
                finally:
                    local_conn.close()
            except Exception:
                pass  # nunca matar el loop por una falla puntual
            _t.sleep(max(60, interval_min * 60))

    t = threading.Thread(target=_worker, daemon=True, name='shopify-ingresos-sync-loop')
    t.start()


# Arrancar el loop al importar el blueprint (una vez por proceso)
try:
    _start_shopify_ingresos_background_loop()
except Exception:
    pass


@bp.route('/api/financiero/sync-shopify-ingresos', methods=['POST'])
def financiero_sync_shopify():
    """Sincroniza pedidos de Shopify NO marcados como flujo_synced
    insertando filas en flujo_ingresos y marcando la orden con
    flujo_synced=1 + flujo_ingreso_id (idempotente y reversible).

    Tambien corre automaticamente:
      - cada vez que se cargan los KPIs del Financiero
      - en background cada SHOPIFY_INGRESOS_SYNC_INTERVAL_MIN minutos

    Body opcional:
      - solo_pagados: bool (default True) — solo importa orders con
        estado_pago='paid'/'pagado'. Evita registrar reservas o
        ordenes cancelados.
      - desde_fecha: 'YYYY-MM-DD' (opcional) — limita el rango.
      - dry_run: bool (default False) — solo cuenta, no escribe.
    """
    if 'compras_user' not in session or session.get('compras_user','') not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.get_json() or {}
    conn = get_db()
    res = _sync_shopify_a_flujo_ingresos(
        conn,
        solo_pagados=d.get('solo_pagados', True),
        desde_fecha=(d.get('desde_fecha') or '').strip(),
        dry_run=bool(d.get('dry_run', False)),
    )
    if not res.get('ok'):
        return jsonify({'error': res.get('error', 'sync fallo'), **res}), 500
    importadas = res.get('importadas', 0)
    total_importado = res.get('total_importado', 0)
    res['mensaje'] = (
        f'{importadas} pedidos Shopify sincronizados a flujo_ingresos. '
        f'Total: ${total_importado:,.0f} COP'
    ) if importadas else 'No hay pedidos Shopify pendientes de sincronizar'
    return jsonify(res)


@bp.route('/api/financiero/sync-shopify-status', methods=['GET'])
def financiero_sync_shopify_status():
    """Estado del sync: cuantos pedidos pendientes hay sin sincronizar."""
    if 'compras_user' not in session or session.get('compras_user','') not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    try:
        pend = c.execute("""SELECT COUNT(*) as n, COALESCE(SUM(total),0) as total
                            FROM animus_shopify_orders
                            WHERE (flujo_synced=0 OR flujo_synced IS NULL)
                              AND LOWER(COALESCE(estado_pago,'')) IN
                                  ('paid','pagado','complete','captured','partially_paid')
                         """).fetchone()
        synced = c.execute("""SELECT COUNT(*) as n, COALESCE(SUM(total),0) as total
                              FROM animus_shopify_orders WHERE flujo_synced=1
                           """).fetchone()
        ultimo_sync = c.execute("""SELECT MAX(fecha) FROM flujo_ingresos
                                   WHERE fuente='shopify_auto'""").fetchone()
        return jsonify({
            'pendientes_count': pend['n'] if pend else 0,
            'pendientes_total': pend['total'] if pend else 0,
            'sincronizados_count': synced['n'] if synced else 0,
            'sincronizados_total': synced['total'] if synced else 0,
            'ultimo_sync_fecha': ultimo_sync[0] if ultimo_sync else None,
        })
    except sqlite3.OperationalError as e:
        return jsonify({'error': f'Tabla no migrada: {e}'}), 500


@bp.route('/api/financiero/limpiar-flujo', methods=['POST'])
def financiero_limpiar_flujo():
    """Elimina todos los registros de flujo_egresos y flujo_ingresos.
    Solo admin. Util para borrar datos de prueba o importaciones erroneas.
    """
    u = session.get('compras_user','')
    if 'compras_user' not in session or u not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json() or {}
    # Requiere confirmacion explicita en el body para evitar borrados accidentales
    if data.get('confirmar') != 'LIMPIAR_TODO':
        return jsonify({'error': 'Falta confirmacion. Envia {"confirmar":"LIMPIAR_TODO"}'}), 400
    conn = get_db(); c = conn.cursor()
    try:
        egr_count = c.execute('SELECT COUNT(*) FROM flujo_egresos').fetchone()[0]
        ing_count = c.execute('SELECT COUNT(*) FROM flujo_ingresos').fetchone()[0]
        # RESPALDO ANTES de borrar (Sebastian 7-ago).
        #
        # Este boton vacia los DOS libros. Esta bien protegido y existe a proposito, pero hasta
        # hoy no tenia vuelta atras: guardaba CUANTAS filas borro, no CUALES. Un boton
        # irreversible sobre la contabilidad funciona bien mil veces, y el dia que alguien se
        # equivoca no hay como deshacerlo.
        #
        # La fila se guarda COMPLETA en JSON, no columna por columna: asi el respaldo sigue
        # sirviendo si manana la tabla gana una columna. Un respaldo con esquema fijo se queda
        # viejo justo cuando hay que usarlo.
        import json as _js_r
        from datetime import datetime as _dt_r, timedelta as _td_r
        _ahora = (_dt_r.utcnow() - _td_r(hours=5)).isoformat(timespec='seconds')
        lote = 'LIMPIEZA-' + _ahora.replace(':', '').replace('-', '')
        for _tabla in ('flujo_egresos', 'flujo_ingresos'):
            _cur = c.execute('SELECT * FROM ' + _tabla)
            _cols = [d[0] for d in _cur.description]
            for _f in _cur.fetchall():
                c.execute(
                    "INSERT INTO flujo_respaldo (lote, tabla, fila_json, creado_en, creado_por) "
                    "VALUES (?,?,?,?,?)",
                    (lote, _tabla, _js_r.dumps(dict(zip(_cols, list(_f))), default=str),
                     _ahora, u))
        # Si el respaldo no guardo todo lo que hay, NO se borra. Un respaldo a medias es peor que
        # ninguno: da la sensacion de que se puede volver atras (M134).
        #
        # Se cuenta LO QUE QUEDO EN LA TABLA, no cuantas veces se intento insertar. La primera
        # version llevaba un contador en el loop y por eso no mordia: contaba la intencion, no el
        # hecho -- si un INSERT no entra, el contador sube igual y el guard aprueba un respaldo
        # que no existe (M5: el numero que decide tiene que ser el medido).
        _guardadas = c.execute("SELECT COUNT(*) FROM flujo_respaldo WHERE lote=?",
                               (lote,)).fetchone()[0] or 0
        if _guardadas != (egr_count + ing_count):
            conn.rollback()
            return jsonify({'error': 'No pude respaldar todo antes de borrar, no se borro nada',
                            'esperadas': egr_count + ing_count,
                            'respaldadas': _guardadas}), 500
        c.execute('DELETE FROM flujo_egresos')
        c.execute('DELETE FROM flujo_ingresos')
        # Audit log critico · borrado masivo de datos financieros
        audit_log(c, usuario=u, accion='LIMPIAR_FLUJO',
                  tabla='flujo_egresos+ingresos', registro_id=None,
                  antes={'egresos_count': egr_count, 'ingresos_count': ing_count},
                  detalle=f"Borrado masivo · {egr_count} egresos + {ing_count} ingresos")
        conn.commit()
        return jsonify({
            'ok': True,
            'eliminados': {'egresos': egr_count, 'ingresos': ing_count},
            'respaldo_lote': lote,
            'message': (f'Limpieza completa: {egr_count} egresos y {ing_count} ingresos '
                        f'eliminados. Respaldados en el lote {lote}: se puede deshacer.'),
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        pass  # conexión cerrada automáticamente por teardown_appcontext

@bp.route('/api/financiero/respaldos-flujo', methods=['GET'])
def financiero_respaldos_flujo():
    """Que respaldos hay para deshacer.

    Va junto con el respaldo, no despues: un respaldo que existe y que nadie sabe que existe no
    sirve el dia que hace falta (M121).
    """
    u = session.get('compras_user', '')
    if 'compras_user' not in session or u not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    c = get_db().cursor()
    try:
        filas = c.execute(
            "SELECT lote, MIN(creado_en), MIN(creado_por), COUNT(*), "
            "       SUM(CASE WHEN COALESCE(restaurado_en,'')<>'' THEN 1 ELSE 0 END) "
            "  FROM flujo_respaldo GROUP BY lote ORDER BY lote DESC LIMIT 50").fetchall()
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    return jsonify({'ok': True, 'lotes': [
        {'lote': r[0], 'cuando': r[1], 'quien': r[2], 'filas': r[3],
         'ya_restaurado': bool(r[4])} for r in filas]})


@bp.route('/api/financiero/restaurar-flujo', methods=['POST'])
def financiero_restaurar_flujo():
    """Deshace una limpieza: vuelve a insertar las filas respaldadas.

    Es lo que convierte el boton de limpiar en una accion reversible. Al restaurar NO se borra el
    respaldo, se MARCA: si alguien restaura por error, el rastro sigue ahi. Es la misma regla que
    el resto del sistema con un registro ya usado (M31/M126).
    """
    u = session.get('compras_user', '')
    if 'compras_user' not in session or u not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.get_json() or {}
    lote = (d.get('lote') or '').strip()
    if not lote:
        return jsonify({'error': 'Falta el lote a restaurar'}), 400
    conn = get_db(); c = conn.cursor()
    try:
        filas = c.execute(
            "SELECT id, tabla, fila_json FROM flujo_respaldo "
            " WHERE lote=? AND COALESCE(restaurado_en,'')='' ORDER BY id", (lote,)).fetchall()
        if not filas:
            # Ni un cero mudo ni un 200 optimista: "no existe" y "ya se restauro" son cosas
            # distintas y llevan a acciones distintas (M100).
            _existe = c.execute("SELECT COUNT(*) FROM flujo_respaldo WHERE lote=?",
                                (lote,)).fetchone()[0]
            return jsonify({'error': ('Ese lote ya se restauro' if _existe
                                      else 'No existe ese lote de respaldo')}), 404
        import json as _js
        from datetime import datetime as _dt2, timedelta as _td2
        _ahora = (_dt2.utcnow() - _td2(hours=5)).isoformat(timespec='seconds')
        n_ok = 0
        for _id, _tabla, _json in filas:
            # La tabla destino se valida contra una lista, no se toma del respaldo: si alguien
            # lograra escribir ahi, un INSERT con nombre de tabla libre seria una puerta abierta.
            if _tabla not in ('flujo_egresos', 'flujo_ingresos'):
                continue
            fila = _js.loads(_json)
            # El id lo asigna la tabla: reusar el viejo chocaria con lo que ya exista.
            fila.pop('id', None)
            cols = list(fila.keys())
            c.execute("INSERT INTO %s (%s) VALUES (%s)"
                      % (_tabla, ','.join(cols), ','.join(['?'] * len(cols))),
                      [fila[k] for k in cols])
            c.execute("UPDATE flujo_respaldo SET restaurado_en=? WHERE id=?", (_ahora, _id))
            n_ok += 1
        audit_log(c, usuario=u, accion='RESTAURAR_FLUJO', tabla='flujo_egresos+ingresos',
                  registro_id=lote, despues={'filas': n_ok},
                  detalle='Restauradas %d filas del lote %s' % (n_ok, lote))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    return jsonify({'ok': True, 'restauradas': n_ok, 'lote': lote})


@bp.route('/api/financiero/precios-mayorista', methods=['GET'])
def get_precios_mayorista():
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT sku, descripcion, precio_base, precio_mayorista, unidad FROM sku_precios ORDER BY sku")
    rows = c.fetchall()
    return jsonify([{'sku':r[0],'descripcion':r[1],'precio_base':r[2],'precio_mayorista':r[3],'unidad':r[4]} for r in rows])

@bp.route('/api/financiero/precios-mayorista/<sku>', methods=['POST'])
def update_precio_mayorista(sku):
    usuario = session.get('compras_user','')
    if 'compras_user' not in session or usuario not in ADMIN_USERS:
        return jsonify({'error': 'Solo admins pueden editar precios'}), 401
    d = request.get_json() or {}
    precio_n, err = validate_money(d.get('precio_mayorista'), allow_zero=True, field_name='precio_mayorista')
    if err:
        return jsonify(err), 400
    conn = get_db(); c = conn.cursor()
    antes = c.execute("SELECT precio_mayorista FROM sku_precios WHERE sku=?", (sku,)).fetchone()
    precio_anterior = float(antes[0]) if antes and antes[0] is not None else None
    c.execute("UPDATE sku_precios SET precio_mayorista=? WHERE sku=?", (precio_n, sku))
    audit_log(c, usuario=usuario, accion='ACTUALIZAR_PRECIO_MAYORISTA',
              tabla='sku_precios', registro_id=sku,
              antes={'precio_mayorista': precio_anterior},
              despues={'precio_mayorista': precio_n},
              detalle=f"SKU {sku}: ${precio_anterior or 0:,.0f} → ${precio_n:,.0f}")
    conn.commit()
    return jsonify({'message': f'Precio actualizado para {sku}'})

@bp.route('/api/financiero/ar-aging')
def financiero_ar_aging():
    from datetime import date
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    today = _hoy_col()
    conn = get_db(); c = conn.cursor()
    # FIX 10-jun audit · la tabla `pedidos` tiene `numero` y `cliente_id` (no
    # `numero_pedido`/`cliente`) → la query daba 500 siempre y la cartera AR nunca
    # cargaba. JOIN a clientes para el nombre (mismo patrón que clientes.py/gerencia.py).
    c.execute("""SELECT p.numero, COALESCE(cl.nombre,''), p.fecha, p.valor_total
                 FROM pedidos p
                 LEFT JOIN clientes cl ON p.cliente_id = cl.id
                 WHERE p.estado NOT IN ('Cancelado','Facturado','Entregado')
                 AND p.valor_total > 0""")
    rows = c.fetchall()
    buckets = {
        'corriente': {'total': 0, 'count': 0},
        'dias_30':   {'total': 0, 'count': 0},
        'dias_60':   {'total': 0, 'count': 0},
        'dias_90':   {'total': 0, 'count': 0},
    }
    pedidos = []
    ar_total = 0
    for r in rows:
        num, cliente, fecha_str, valor = r
        try:
            fd = date.fromisoformat(fecha_str[:10])
        except Exception:
            fd = today
        dias = (today - fd).days
        ar_total += (valor or 0)
        if dias <= 30:
            b = 'corriente'
        elif dias <= 60:
            b = 'dias_30'
        elif dias <= 90:
            b = 'dias_60'
        else:
            b = 'dias_90'
        buckets[b]['total'] += (valor or 0)
        buckets[b]['count'] += 1
        pedidos.append({'numero_pedido': num, 'cliente': cliente, 'fecha': fecha_str[:10] if fecha_str else '', 'dias': dias, 'valor_total': valor or 0})
    pedidos.sort(key=lambda x: x['dias'], reverse=True)
    return jsonify({'ar_total': ar_total, 'count': len(pedidos), 'buckets': buckets, 'pedidos': pedidos})

@bp.route('/api/financiero/ap-aging')
def financiero_ap_aging():
    from datetime import date
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    today = _hoy_col()
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT numero_oc, proveedor, fecha, valor_total
                 FROM ordenes_compra
                 WHERE estado IN ('Autorizada','Recibida','Parcial')
                 AND (pagado_por IS NULL OR pagado_por = '')""")
    rows = c.fetchall()
    buckets = {
        'corriente': {'total': 0, 'count': 0},
        'dias_30':   {'total': 0, 'count': 0},
        'dias_60':   {'total': 0, 'count': 0},
        'dias_90':   {'total': 0, 'count': 0},
    }
    ocs = []
    ap_total = 0
    for r in rows:
        num, prov, fecha_str, valor = r
        try:
            fd = date.fromisoformat(fecha_str[:10])
        except Exception:
            fd = today
        dias = (today - fd).days
        ap_total += (valor or 0)
        if dias <= 30:
            b = 'corriente'
        elif dias <= 60:
            b = 'dias_30'
        elif dias <= 90:
            b = 'dias_60'
        else:
            b = 'dias_90'
        buckets[b]['total'] += (valor or 0)
        buckets[b]['count'] += 1
        ocs.append({'numero_oc': num, 'proveedor': prov, 'fecha': fecha_str[:10] if fecha_str else '', 'dias': dias, 'valor_total': valor or 0})
    ocs.sort(key=lambda x: x['dias'], reverse=True)
    return jsonify({'ap_total': ap_total, 'count': len(ocs), 'buckets': buckets, 'ocs': ocs})

@bp.route('/api/financiero/working-capital')
def financiero_working_capital():
    from datetime import date, timedelta
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    today = _hoy_col()
    conn = get_db(); c = conn.cursor()
    # AR
    c.execute("SELECT COALESCE(SUM(valor_total),0) FROM pedidos WHERE estado NOT IN ('Cancelado','Facturado','Entregado') AND valor_total > 0")
    ar_total = c.fetchone()[0] or 0
    # AP
    c.execute("SELECT COALESCE(SUM(valor_total),0) FROM ordenes_compra WHERE estado IN ('Autorizada','Recibida','Parcial') AND (pagado_por IS NULL OR pagado_por='')")
    ap_total = c.fetchone()[0] or 0
    # Cash from gerencia_inputs
    try:
        # FIX 10-jun audit · gerencia_inputs es (periodo, saldo_caja, ...) · NO tiene
        # columnas clave/valor → la query fallaba siempre y dejaba cash=0 (working_capital
        # y runway subestimados). Misma fuente que financiero_kpis y gerencia.
        c.execute("SELECT saldo_caja FROM gerencia_inputs ORDER BY periodo DESC LIMIT 1")
        row = c.fetchone()
        cash = float(row[0]) if row else 0.0
    except Exception:
        cash = 0.0
    # Valor del inventario de MP · sale del kardex canonico (SUM(movimientos)
    # excluyendo los 6 estados no disponibles) con el precio en $/g.
    # 19-ago · se retira la consulta que leia la tabla `lotes`, que NO EXISTE en este
    # esquema: fallaba siempre y el `except` la dejaba en 0, asi que este camino era el
    # que decidia igual. El numero mostrado ya era correcto; lo que sobraba era pagar
    # una consulta imposible en cada carga (y su rollback en PostgreSQL).
    try:
        _filas = c.execute(
            "SELECT m.material_id, "
            " SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad "
            "          WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END), "
            " COALESCE(NULLIF((SELECT mp.precio_referencia FROM maestro_mps mp "
            "                   WHERE mp.codigo_mp=m.material_id),0)/1000.0, "
            "          (SELECT AVG(oci.precio_unitario) FROM ordenes_compra_items oci "
            "            WHERE oci.codigo_mp=m.material_id AND oci.precio_unitario>0), 0) "
            "FROM movimientos m "
            "WHERE UPPER(COALESCE(m.estado_lote,'')) NOT IN "
            "      ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO') "
            "GROUP BY m.material_id").fetchall()
        inventory_value = sum(float(r[1] or 0) * float(r[2] or 0)
                              for r in _filas if float(r[1] or 0) > 0.01)
    except Exception as _e_inv:
        import logging as _lg_inv
        _lg_inv.getLogger('financiero').warning(
            'valor de inventario no se pudo calcular: %s', _e_inv)
        inventory_value = 0.0
    # 90-day flows for DSO/DIO/DPO
    cutoff90 = (today - timedelta(days=90)).isoformat()
    c.execute("SELECT COALESCE(SUM(valor_total),0) FROM pedidos WHERE fecha >= ? AND estado NOT IN ('Cancelado')", (cutoff90,))
    ventas_90 = c.fetchone()[0] or 1
    c.execute("SELECT COALESCE(SUM(valor_total),0) FROM ordenes_compra WHERE fecha >= ? AND estado NOT IN ('Pendiente','Cancelada')", (cutoff90,))
    compras_90 = c.fetchone()[0] or 1
    c.execute("SELECT COALESCE(SUM(fi.monto),0) FROM flujo_egresos fi WHERE fi.fecha >= ? AND fi.categoria IN ('MP','MPs','Materia Prima','Insumo')", (cutoff90,))
    cogs_90 = c.fetchone()[0] or 1
    # Burn rate: promedio mensual de OCs pagadas (últimos 3 meses)
    cutoff3m = (today - timedelta(days=90)).isoformat()
    c.execute("SELECT COALESCE(SUM(valor_total),0) FROM ordenes_compra "
              "WHERE fecha >= ? AND estado NOT IN ('Pendiente','Cancelada','Borrador')",
              (cutoff3m,))
    egr3m = c.fetchone()[0] or 0
    burn_rate = max(egr3m / 3.0, 1.0)
    dso = (ar_total / (ventas_90 / 90.0)) if ventas_90 > 0 else 0
    dpo = (ap_total / (compras_90 / 90.0)) if compras_90 > 0 else 0
    dio = (inventory_value / (cogs_90 / 90.0)) if cogs_90 > 0 else 0
    ccc = dio + dso - dpo
    working_capital = cash + inventory_value + ar_total - ap_total
    runway_meses = (cash / burn_rate) if burn_rate > 0 else 0
    return jsonify({
        'ar_total': ar_total, 'ap_total': ap_total, 'cash': cash,
        'inventory_value': inventory_value, 'working_capital': working_capital,
        'dso': dso, 'dpo': dpo, 'dio': dio, 'ccc': ccc,
        'burn_rate': burn_rate, 'runway_meses': runway_meses
    })

@bp.route('/api/financiero/pnl')
def financiero_pnl():
    """P&L real: ingresos desde pedidos + maquila, egresos desde ordenes_compra."""
    from datetime import date, timedelta
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    today   = _hoy_col()
    mes_str = today.strftime('%Y-%m')
    year_str= today.strftime('%Y')
    periodo = today.strftime('%b %Y')
    conn = get_db(); c = conn.cursor()

    def ing_animus(periodo_like):
        c.execute("SELECT COALESCE(SUM(valor_total),0) FROM pedidos "
                  "WHERE fecha LIKE ? AND estado NOT IN ('Cancelado')"
                  " AND (empresa='ANIMUS' OR empresa IS NULL OR empresa='')",
                  (periodo_like+'%',))
        return c.fetchone()[0] or 0

    def ing_maquila(periodo_like):
        try:
            c.execute("SELECT COALESCE(SUM(precio_lote),0) FROM maquila_ordenes "
                      "WHERE fecha_orden LIKE ? AND estado NOT IN ('Cotizacion','Cancelada')",
                      (periodo_like+'%',))
            return c.fetchone()[0] or 0
        except Exception:
            return 0

    def egr_total(periodo_like):
        c.execute("SELECT COALESCE(SUM(valor_total),0) FROM ordenes_compra "
                  "WHERE fecha LIKE ? AND estado NOT IN ('Pendiente','Cancelada','Borrador')",
                  (periodo_like+'%',))
        return c.fetchone()[0] or 0

    # Mes actual
    animus_ing  = ing_animus(mes_str)
    maqui_ing   = ing_maquila(mes_str)
    total_ing   = animus_ing + maqui_ing
    total_egr   = egr_total(mes_str)
    margen      = total_ing - total_egr
    margen_pct  = round((margen / total_ing * 100), 1) if total_ing > 0 else 0
    # YTD
    ytd_ing = ing_animus(year_str) + ing_maquila(year_str)
    ytd_egr = egr_total(year_str)
    empresas = {
        'ANIMUS':    {'ingresos': animus_ing, 'egresos': 0, 'margen': animus_ing,
                      'margen_pct': 100, 'ingresos_ytd': ing_animus(year_str),
                      'egresos_ytd': 0, 'ebitda': animus_ing},
        'ESPAGIRIA': {'ingresos': maqui_ing, 'egresos': 0, 'margen': maqui_ing,
                      'margen_pct': 100, 'ingresos_ytd': ing_maquila(year_str),
                      'egresos_ytd': 0, 'ebitda': maqui_ing},
        'TOTAL':     {'ingresos': total_ing, 'egresos': total_egr, 'margen': margen,
                      'margen_pct': margen_pct, 'ingresos_ytd': ytd_ing,
                      'egresos_ytd': ytd_egr, 'ebitda': margen},
    }
    # Historico 6 meses
    historico = []
    for i in range(5, -1, -1):
        ref   = today.replace(day=1) - timedelta(days=i * 28)
        p     = ref.strftime('%Y-%m')
        label = ref.strftime('%b %y')
        h_ing = ing_animus(p) + ing_maquila(p)
        h_egr = egr_total(p)
        historico.append({'periodo': label, 'ingresos': h_ing,
                          'egresos': h_egr, 'margen': h_ing - h_egr})
    # YTD vs anio anterior
    prev_year = str(int(year_str) - 1)
    ytd_prev_ing = ing_animus(prev_year) + ing_maquila(prev_year)
    ytd_prev_egr = egr_total(prev_year)
    ytd_crecimiento = round((ytd_ing - ytd_prev_ing) / ytd_prev_ing * 100, 1) if ytd_prev_ing > 0 else None
    # Mes actual vs mismo mes anio anterior
    mes_prev = today.replace(year=today.year - 1).strftime('%Y-%m')
    prev_mes_ing = ing_animus(mes_prev) + ing_maquila(mes_prev)
    prev_mes_egr = egr_total(mes_prev)
    mes_crecimiento = round((total_ing - prev_mes_ing) / prev_mes_ing * 100, 1) if prev_mes_ing > 0 else None
    # Nomina este mes (para costo laboral en P&L)
    try:
        c.execute("SELECT COALESCE(SUM(salario_neto),0) FROM nomina_registros WHERE periodo=?", (mes_str,))
        nomina_mes = c.fetchone()[0] or 0
    except Exception:
        try:
            c.execute("SELECT COALESCE(SUM(salario_base),0) FROM empleados WHERE estado='Activo'")
            nomina_mes = c.fetchone()[0] or 0
        except Exception:
            nomina_mes = 0
    ebitda = total_ing - total_egr - nomina_mes
    return jsonify({
        'empresas': empresas, 'historico': historico, 'periodo': periodo,
        'ytd': {'ingresos': ytd_ing, 'egresos': ytd_egr, 'margen': ytd_ing - ytd_egr,
                'prev_ingresos': ytd_prev_ing, 'crecimiento_pct': ytd_crecimiento},
        'mes_vs_prior': {'ingresos': total_ing, 'prev_ingresos': prev_mes_ing, 'crecimiento_pct': mes_crecimiento},
        'nomina_mes': nomina_mes,
        'ebitda': ebitda,
    })

# ===============================================================
# INVENTARIO v2 - NUEVOS ENDPOINTS
# ===============================================================


# ══ LIBRO DE ACTIVOS (mig 403) ═════════════════════════════════════════════════
# Sebastián 30-jul: *"esto es trazabilidad y plata, pero a la vez nos permite hacer
# seguimientos · como CEO debo verlo, Tesorería también, y todo lo que llegue se debe
# recepcionar"*. El maestro vivía en un Excel: el valor de la empresa dependía de que
# nadie perdiera un archivo.
#
# MP y envases NO entran acá (decisión suya): son inventario VIVO que varía con el uso y
# ya tienen su kardex. Meterlos duplicaría la verdad (M37).

# Un activo dado de baja, hurtado o fuera de uso DEJA de sumar al valor en libros, pero la
# fila NO se borra: un activo robado no desaparece del libro, se registra como pérdida.
# 'Dañado' SÍ suma: sigue siendo un bien, sólo que deteriorado (el Excel lo marca como
# "para revisión / posible baja", que es una decisión pendiente, no un hecho consumado).
_ACTIVO_FUERA_DE_LIBROS = ('DE BAJA', 'BAJA', 'HURTO', 'ROBADO', 'PERDIDO', 'FUERA DE USO')


def _activo_auth():
    """CEO (admin) y Tesorería (contadora). Mismo criterio que el resto de /financiero."""
    u = session.get('compras_user', '')
    if not u or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return None, jsonify({'error': 'No autorizado'}), 401
    return u, None, None


def _activo_en_libros(estado):
    return str(estado or '').strip().upper() not in _ACTIVO_FUERA_DE_LIBROS


def _activo_valor(costo, deprec):
    v = float(costo or 0) - float(deprec or 0)
    return v if v > 0 else 0.0


@bp.route('/api/activos', methods=['GET'])
def activos_listar():
    """El libro completo + los totales. `valor_en_libros` se DERIVA, no se teclea."""
    _u, _err, _code = _activo_auth()
    if _err:
        return _err, _code
    conn = get_db(); c = conn.cursor()
    try:
        filas = c.execute(
            "SELECT codigo, empresa, nombre, tipo_bien, categoria_contable, ubicacion, "
            "responsable, cantidad, estado, rotulado, costo_cop, vida_util_anios, "
            "fecha_ingreso, depreciacion_acumulada_cop, COALESCE(serial,''), "
            "COALESCE(factura,''), COALESCE(proveedor,''), COALESCE(equipo_codigo,''), "
            "origen, COALESCE(notas,''), COALESCE(baja_motivo,''), COALESCE(baja_fecha,'') "
            "FROM activos ORDER BY empresa, codigo").fetchall()
    except Exception as e:
        return jsonify({'error': 'no se pudo leer el libro', 'detalle': str(e)}), 500
    items, kpis = [], {}
    por_estado, por_categoria = {}, {}
    for r in filas:
        en_libros = _activo_en_libros(r[8])
        valor = _activo_valor(r[10], r[13]) if en_libros else 0.0
        emp = (r[1] or 'SIN EMPRESA').strip()
        items.append({
            'codigo': r[0], 'empresa': emp, 'nombre': r[2], 'tipo_bien': r[3],
            'categoria_contable': r[4], 'ubicacion': r[5], 'responsable': r[6],
            'cantidad': float(r[7] or 1), 'estado': r[8], 'rotulado': int(r[9] or 0),
            'costo_cop': float(r[10] or 0), 'vida_util_anios': float(r[11] or 0),
            'fecha_ingreso': r[12], 'depreciacion_cop': float(r[13] or 0),
            'serial': r[14], 'factura': r[15], 'proveedor': r[16], 'equipo_codigo': r[17],
            'origen': r[18], 'notas': r[19], 'baja_motivo': r[20], 'baja_fecha': r[21],
            'en_libros': en_libros, 'valor_en_libros': round(valor, 2),
        })
        k = kpis.setdefault(emp, {'activos': 0, 'valor': 0.0, 'fuera': 0, 'sin_rotular': 0})
        k['activos'] += 1
        k['valor'] += valor
        if not en_libros:
            k['fuera'] += 1
        if not int(r[9] or 0):
            k['sin_rotular'] += 1
        _e = (r[8] or 'Sin estado').strip()
        por_estado[_e] = por_estado.get(_e, 0) + 1
        _cat = (r[4] or 'Sin categoria').strip()
        _c = por_categoria.setdefault(_cat, {'n': 0, 'valor': 0.0})
        _c['n'] += 1
        _c['valor'] += valor
    total_valor = round(sum(k['valor'] for k in kpis.values()), 2)
    for k in kpis.values():
        k['valor'] = round(k['valor'], 2)
    for v in por_categoria.values():
        v['valor'] = round(v['valor'], 2)
    return jsonify({
        'ok': True, 'items': items, 'total': len(items),
        'valor_en_libros_total': total_valor,
        'por_empresa': kpis, 'por_estado': por_estado, 'por_categoria': por_categoria,
        'estados_fuera_de_libros': sorted(_ACTIVO_FUERA_DE_LIBROS),
    })


@bp.route('/api/activos/importar', methods=['POST'])
def activos_importar():
    """Carga el Excel maestro de activos. `?dry_run=1` (default) NO escribe: muestra el plan.

    Cruza las dos hojas por CODIGO: `INVENTARIO HHA` (qué es, dónde está, quién responde) y
    `CIERRE CONTABLE` (categoría, vida útil, costo, depreciación). Sebastián 30-jul: *"el Excel
    manda, sólo usá los que te subo"* -- así que esto no inventa activos ni borra los que no
    vengan: da de alta y ACTUALIZA, y lo que sobra queda listado para que él decida.
    """
    _u, _err, _code = _activo_auth()
    if _err:
        return _err, _code
    f = request.files.get('archivo')
    if not f:
        return jsonify({'error': 'Subí el Excel maestro de activos (campo `archivo`)'}), 400
    dry = str(request.args.get('dry_run', '1')).strip() not in ('0', 'false', 'no')
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl no está disponible en el servidor'}), 500
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({'error': 'no se pudo leer el archivo', 'detalle': str(e)}), 400
    if 'INVENTARIO HHA' not in wb.sheetnames:
        return jsonify({'error': 'el archivo no tiene la hoja "INVENTARIO HHA"',
                        'hojas': wb.sheetnames}), 400

    def _txt(x):
        return str(x).strip() if x is not None else ''

    def _num(x):
        try:
            return float(str(x).replace(',', '').replace('$', '').strip() or 0)
        except (TypeError, ValueError):
            return 0.0

    inv = {}
    incompletos = []
    for r in wb['INVENTARIO HHA'].iter_rows(min_row=4, values_only=True):
        cod = _txt(r[0] if len(r) > 0 else '')
        _desc = _txt(r[3] if len(r) > 3 else '')
        _tipo = _txt(r[2] if len(r) > 2 else '')
        if not cod and not _desc and not _tipo:
            continue                      # fila vacía / separador
        # El nombre cae al TIPO DE BIEN si la descripción viene vacía. Tres activos reales del
        # archivo (una balanza y dos tapadoras) traen código y tipo pero sin descripción: con el
        # filtro original se perdían SIN AVISO, que es la peor forma de perder un activo. Se
        # cargan igual y se DECLARAN en el plan para que se corrija el Excel.
        _nombre = _desc or _tipo or cod
        if not cod or not _desc:
            incompletos.append({'codigo': cod or '(sin código)', 'nombre': _nombre,
                                'falta': ('código' if not cod else 'descripción')})
        if not cod:
            continue                      # sin código no hay llave: no se puede cargar
        inv[cod] = {
            'codigo': cod, 'empresa': _txt(r[1]), 'tipo_bien': _tipo,
            'nombre': _nombre, 'ubicacion': _txt(r[4]), 'responsable': _txt(r[5]),
            'cantidad': _num(r[6]) or 1, 'estado': _txt(r[7]) or 'En uso',
            'rotulado': 1 if _txt(r[8]).upper().startswith('S') else 0,
            'costo_cop': _num(r[9]),
            'categoria_contable': '', 'vida_util_anios': 0.0, 'fecha_ingreso': '',
            'depreciacion_acumulada_cop': 0.0, 'notas': '',
        }
    if 'CIERRE CONTABLE' in wb.sheetnames:
        for r in wb['CIERRE CONTABLE'].iter_rows(min_row=4, values_only=True):
            cod = _txt(r[0] if len(r) > 0 else '')
            if cod not in inv:
                continue
            inv[cod].update({
                'categoria_contable': _txt(r[3]), 'vida_util_anios': _num(r[4]),
                'fecha_ingreso': _txt(r[5])[:10],
                'costo_cop': _num(r[7]) or inv[cod]['costo_cop'],
                'depreciacion_acumulada_cop': _num(r[8]),
                'notas': _txt(r[10]),
            })
    if not inv:
        return jsonify({'error': 'no se encontró ninguna fila de activo en el archivo'}), 400

    conn = get_db(); c = conn.cursor()
    existentes = {}
    try:
        for r in c.execute("SELECT codigo, nombre, estado, costo_cop FROM activos").fetchall():
            existentes[r[0]] = {'nombre': r[1], 'estado': r[2], 'costo': float(r[3] or 0)}
    except Exception as e:
        return jsonify({'error': 'no se pudo leer el libro actual', 'detalle': str(e)}), 500
    nuevos = [k for k in inv if k not in existentes]
    cambios = []
    for k, v in inv.items():
        e = existentes.get(k)
        if not e:
            continue
        if (e['nombre'] != v['nombre'] or e['estado'] != v['estado']
                or abs(e['costo'] - v.get('costo_cop', 0)) > 0.01):
            cambios.append({'codigo': k, 'antes': e,
                            'despues': {'nombre': v['nombre'], 'estado': v['estado'],
                                        'costo': v.get('costo_cop', 0)}})
    sobran = [k for k in existentes if k not in inv]
    plan = {
        'ok': True, 'dry_run': dry, 'en_archivo': len(inv),
        'nuevos': len(nuevos), 'actualizan': len(cambios), 'no_vienen_en_el_archivo': len(sobran),
        'detalle_nuevos': nuevos[:40], 'detalle_cambios': cambios[:40],
        'detalle_sobran': sobran[:40],
        'incompletos': incompletos[:40], 'n_incompletos': len(incompletos),
        'aviso': ('Los %d que ya están en EOS y NO vienen en el archivo NO se tocan: el import '
                  'da de alta y actualiza, nunca borra. Si alguno hay que darlo de baja, se hace '
                  'con su motivo desde la pantalla.' % len(sobran)) if sobran else '',
        'aviso_incompletos': ('%d fila(s) del archivo vienen sin descripción (se cargan con el '
                              'tipo de bien como nombre): %s. Conviene completarlas en el Excel.'
                              % (len(incompletos),
                                 ', '.join(x['codigo'] for x in incompletos[:8]))) if incompletos else '',
    }
    if dry:
        return jsonify(plan)

    ahora = datetime.utcnow().replace(microsecond=0).isoformat()
    for k, v in inv.items():
        campos = (v.get('empresa', ''), v.get('nombre', ''), v.get('tipo_bien', ''),
                  v.get('categoria_contable', ''), v.get('ubicacion', ''),
                  v.get('responsable', ''), float(v.get('cantidad') or 1),
                  v.get('estado', 'En uso'), int(v.get('rotulado') or 0),
                  float(v.get('costo_cop') or 0), float(v.get('vida_util_anios') or 0),
                  v.get('fecha_ingreso', ''),
                  float(v.get('depreciacion_acumulada_cop') or 0), v.get('notas', ''), ahora)
        if k in existentes:
            c.execute(
                "UPDATE activos SET empresa=?, nombre=?, tipo_bien=?, categoria_contable=?, "
                "ubicacion=?, responsable=?, cantidad=?, estado=?, rotulado=?, costo_cop=?, "
                "vida_util_anios=?, fecha_ingreso=?, depreciacion_acumulada_cop=?, notas=?, "
                "actualizado_en=? WHERE codigo=?", campos + (k,))
        else:
            c.execute(
                "INSERT INTO activos (empresa, nombre, tipo_bien, categoria_contable, ubicacion, "
                "responsable, cantidad, estado, rotulado, costo_cop, vida_util_anios, "
                "fecha_ingreso, depreciacion_acumulada_cop, notas, actualizado_en, codigo, origen) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'excel')", campos + (k,))
            c.execute("INSERT INTO activos_eventos (activo_codigo, tipo, detalle, "
                      "valor_despues, estado_despues, usuario) VALUES (?,'ALTA',?,?,?,?)",
                      (k, 'Alta por importacion del Excel maestro',
                       float(v.get('costo_cop') or 0), v.get('estado', 'En uso'), _u))
    audit_log(c, usuario=_u, accion='ACTIVOS_IMPORTAR', tabla='activos',
              registro_id='excel',
              despues={'en_archivo': len(inv), 'nuevos': len(nuevos), 'actualizan': len(cambios)},
              detalle='Import del Excel maestro de activos · %d fila(s)' % len(inv))
    conn.commit()
    plan['dry_run'] = False
    plan['aplicado'] = True
    return jsonify(plan)


@bp.route('/api/activos/<path:codigo>/baja', methods=['POST'])
def activos_baja(codigo):
    """Da de baja un activo (o revierte la baja). Sale del valor en libros SIN borrarse.

    Body: `{motivo, estado}` con estado en De baja / Hurto / Fuera de uso, o `{revertir: true}`.
    """
    _u, _err, _code = _activo_auth()
    if _err:
        return _err, _code
    d = request.get_json(silent=True) or {}
    cod = str(codigo or '').strip()
    conn = get_db(); c = conn.cursor()
    fila = c.execute("SELECT estado, costo_cop, depreciacion_acumulada_cop, nombre "
                     "FROM activos WHERE codigo=?", (cod,)).fetchone()
    if not fila:
        return jsonify({'error': 'activo %s no existe' % cod}), 404
    hoy = _hoy_col().isoformat()
    ahora = datetime.utcnow().replace(microsecond=0).isoformat()
    if d.get('revertir'):
        nuevo = str(d.get('estado') or 'En uso').strip() or 'En uso'
        c.execute("UPDATE activos SET estado=?, baja_motivo='', baja_fecha='', baja_por='', "
                  "actualizado_en=? WHERE codigo=?", (nuevo, ahora, cod))
        c.execute("INSERT INTO activos_eventos (activo_codigo, tipo, detalle, estado_antes, "
                  "estado_despues, usuario) VALUES (?,'REVERTIR_BAJA',?,?,?,?)",
                  (cod, str(d.get('motivo') or '')[:300], fila[0], nuevo, _u))
        audit_log(c, usuario=_u, accion='ACTIVO_REVERTIR_BAJA', tabla='activos', registro_id=cod,
                  antes={'estado': fila[0]}, despues={'estado': nuevo},
                  detalle='Revierte la baja de %s' % cod)
        conn.commit()
        return jsonify({'ok': True, 'codigo': cod, 'estado': nuevo, 'en_libros': True})
    motivo = str(d.get('motivo') or '').strip()
    if not motivo:
        # Una baja sin motivo es plata que desaparece del libro sin explicación.
        return jsonify({'error': 'La baja va con motivo: es plata que sale del valor en libros'}), 400
    estado = str(d.get('estado') or 'De baja').strip()
    if estado.upper() not in _ACTIVO_FUERA_DE_LIBROS:
        return jsonify({'error': 'estado inválido para una baja',
                        'validos': sorted(_ACTIVO_FUERA_DE_LIBROS)}), 400
    if not _activo_en_libros(fila[0]):
        return jsonify({'error': 'ese activo ya estaba fuera de libros (%s)' % fila[0],
                        'codigo_error': 'YA_DE_BAJA'}), 409
    valor_antes = _activo_valor(fila[1], fila[2])
    # CAS: dos bajas concurrentes no pueden registrar dos pérdidas del mismo activo (M27).
    c.execute("UPDATE activos SET estado=?, baja_motivo=?, baja_fecha=?, baja_por=?, "
              "actualizado_en=? WHERE codigo=? AND COALESCE(baja_fecha,'')=''",
              (estado, motivo[:300], hoy, _u, ahora, cod))
    if c.rowcount != 1:
        conn.rollback()
        return jsonify({'error': 'otro usuario ya dio de baja este activo · refrescá',
                        'codigo_error': 'YA_DE_BAJA'}), 409
    c.execute("INSERT INTO activos_eventos (activo_codigo, tipo, detalle, valor_antes, "
              "valor_despues, estado_antes, estado_despues, usuario) "
              "VALUES (?,'BAJA',?,?,0,?,?,?)", (cod, motivo[:300], valor_antes, fila[0], estado, _u))
    audit_log(c, usuario=_u, accion='ACTIVO_BAJA', tabla='activos', registro_id=cod,
              antes={'estado': fila[0], 'valor_en_libros': valor_antes},
              despues={'estado': estado, 'valor_en_libros': 0, 'motivo': motivo},
              detalle='Baja de %s (%s) · %s' % (cod, fila[3], estado))
    conn.commit()
    return jsonify({'ok': True, 'codigo': cod, 'estado': estado, 'en_libros': False,
                    'valor_que_sale': round(valor_antes, 2)})


@bp.route('/activos', methods=['GET'])
def activos_page():
    u = session.get('compras_user', '')
    if 'compras_user' not in session or (u not in ADMIN_USERS and u not in CONTADORA_USERS):
        return redirect(url_for('core.login'))
    from templates_py.activos_html import ACTIVOS_HTML
    return Response(ACTIVOS_HTML, mimetype='text/html')
