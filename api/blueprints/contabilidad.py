# blueprints/contabilidad.py — Centro de mando de la contadora (Mayra)
import json
import sqlite3
from datetime import datetime, date, timedelta
from io import BytesIO

from flask import Blueprint, jsonify, request, Response, session
from config import DB_PATH, ADMIN_USERS, CONTADORA_USERS
from database import get_db

bp = Blueprint('contabilidad', __name__)
CONT_USERS = CONTADORA_USERS | ADMIN_USERS


# ─── Auth helpers ─────────────────────────────────────────────────────────────
def _auth():
    # Acepta sesion propia del modulo O la sesion principal del sistema
    return (
        session.get('cont_user', '') in CONT_USERS
        or session.get('compras_user', '') in CONT_USERS
    )

def _user():
    return session.get('cont_user', '') or session.get('compras_user', 'sistema')


# ─── Numeración secuencial atómica ───────────────────────────────────────────
def _next_numero(conn, empresa: str, tipo: str = 'FV') -> str:
    anio = date.today().year
    prefix = {'ANIMUS': 'ANI', 'Espagiria': 'ESP'}.get(empresa, 'GEN')
    conn.execute(
        "INSERT OR IGNORE INTO config_facturacion(empresa,anio,tipo,siguiente) VALUES(?,?,?,1)",
        (empresa, anio, tipo)
    )
    row = conn.execute(
        "SELECT siguiente FROM config_facturacion WHERE empresa=? AND anio=? AND tipo=?",
        (empresa, anio, tipo)
    ).fetchone()
    seq = row[0] if row else 1
    conn.execute(
        "UPDATE config_facturacion SET siguiente=siguiente+1 WHERE empresa=? AND anio=? AND tipo=?",
        (empresa, anio, tipo)
    )
    conn.commit()
    return f"{tipo}-{prefix}-{anio}-{seq:04d}"


# ─── PDF con fpdf2 ───────────────────────────────────────────────────────────
def _generar_pdf_bytes(factura: dict, items: list) -> bytes:
    from fpdf import FPDF

    class PDF(FPDF):
        def header(self):
            pass
        def footer(self):
            self.set_y(-12)
            self.set_font('Helvetica', 'I', 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 6, f'Documento generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} - Pag. {self.page_no()}', align='C')

    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=18)
    W = pdf.w - 20  # ancho util

    empresa = factura.get('empresa', 'ANIMUS')
    nombre_empresa = 'ANIMUS LAB S.A.S.' if empresa == 'ANIMUS' else 'ESPAGIRIA LABORATORIO S.A.S.'

    # ── Encabezado empresa ────────────────────────────────────────────────────
    pdf.set_fill_color(20, 20, 20)
    pdf.rect(10, 10, W, 22, 'F')
    pdf.set_xy(12, 13)
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(W - 4, 8, nombre_empresa, ln=True)
    pdf.set_x(12)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(200, 200, 200)
    pdf.cell(W - 4, 5, 'Cali, Valle del Cauca, Colombia', ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(6)

    # ── Titulo documento ──────────────────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 13)
    pdf.cell(W, 8, factura.get('tipo', 'Factura').upper() + ' DE VENTA', ln=True, align='C')
    pdf.ln(2)

    # ── Metadatos ─────────────────────────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(240, 240, 240)
    col = W / 3
    pdf.cell(col, 7, f"No: {factura['numero']}", border=1, fill=True)
    pdf.cell(col, 7, f"Fecha: {factura.get('fecha_emision','')}", border=1, fill=True)
    venc = factura.get('fecha_vencimiento', '')
    pdf.cell(col, 7, f"Vence: {venc if venc else 'Contado'}", border=1, fill=True, ln=True)
    if factura.get('numero_pedido'):
        pdf.set_font('Helvetica', '', 8)
        pdf.cell(W, 5, f"Pedido referencia: {factura['numero_pedido']}", border=0, ln=True)
    pdf.ln(4)

    # ── Datos cliente ─────────────────────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(W, 7, ' CLIENTE', border=1, fill=True, ln=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(col * 2, 6, f"  Nombre: {factura.get('cliente_nombre','')}", border=1)
    pdf.cell(col, 6, f"  NIT: {factura.get('cliente_nit','')}", border=1, ln=True)
    pdf.ln(4)

    # ── Tabla de items ────────────────────────────────────────────────────────
    widths = [8, 25, 80, 18, 24, 25]  # cols: #, SKU, Descripcion, Cant, P.Unit, Subtotal
    headers = ['#', 'SKU', 'Descripcion', 'Cant', 'P. Unitario', 'Subtotal']
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(20, 20, 20)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, headers):
        pdf.cell(w, 7, h, border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 8)
    for i, it in enumerate(items, 1):
        fill = i % 2 == 0
        pdf.set_fill_color(248, 248, 248)
        sub = it.get('subtotal', 0)
        vals = [
            str(i),
            it.get('sku', ''),
            it.get('descripcion', ''),
            str(it.get('cantidad', 0)),
            f"${it.get('precio_unitario', 0):,.0f}",
            f"${sub:,.0f}",
        ]
        for w, v in zip(widths, vals):
            pdf.cell(w, 6, v, border=1, fill=fill, align='C' if v.isdigit() or v.startswith('$') else 'L')
        pdf.ln()
    pdf.ln(3)

    # ── Totales ───────────────────────────────────────────────────────────────
    pdf.set_font('Helvetica', '', 9)
    right_x = 10 + W - 60
    sub_t = factura.get('subtotal', 0)
    desc = factura.get('descuento', 0)
    iva_pct = factura.get('iva_pct', 0)
    iva_v = factura.get('iva_valor', 0)
    total = factura.get('total', 0)

    def tot_row(label, val, bold=False):
        pdf.set_x(right_x)
        if bold:
            pdf.set_font('Helvetica', 'B', 10)
        else:
            pdf.set_font('Helvetica', '', 9)
        pdf.cell(35, 6, label, border=0, align='R')
        pdf.cell(25, 6, f"${val:,.0f}", border=1, align='R', ln=True)

    tot_row('Subtotal:', sub_t)
    if desc > 0:
        tot_row(f'Descuento:', desc)
    if iva_pct > 0:
        tot_row(f'IVA ({iva_pct:.0f}%):', iva_v)
    tot_row('TOTAL COP:', total, bold=True)
    pdf.ln(6)

    # ── Notas ────────────────────────────────────────────────────────────────
    if factura.get('notas'):
        pdf.set_font('Helvetica', 'I', 8)
        pdf.set_text_color(80, 80, 80)
        pdf.multi_cell(W, 5, f"Notas: {factura['notas']}")
    pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output())


# ─── Excel Siigo ─────────────────────────────────────────────────────────────
def _generar_excel_siigo(facturas: list, items_map: dict, pagos_map: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()

    # ── Hoja 1: Libro de Ventas ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Libro de Ventas"
    hdr_fill = PatternFill("solid", fgColor="1A1A1A")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers1 = [
        'Tipo', 'Numero', 'Fecha Emision', 'Fecha Vencimiento',
        'NIT Cliente', 'Nombre Cliente', 'Empresa',
        'Subtotal', 'Descuento', 'IVA %', 'IVA Valor', 'Total',
        'Estado', 'Pedido Ref', 'Notas'
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    for row, f in enumerate(facturas, 2):
        vals = [
            f.get('tipo'), f.get('numero'), f.get('fecha_emision'),
            f.get('fecha_vencimiento'), f.get('cliente_nit'), f.get('cliente_nombre'),
            f.get('empresa'), f.get('subtotal'), f.get('descuento'),
            f.get('iva_pct'), f.get('iva_valor'), f.get('total'),
            f.get('estado'), f.get('numero_pedido'), f.get('notas')
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row, col, v)
            cell.border = thin
            if isinstance(v, (int, float)):
                cell.number_format = '#,##0'

    # Auto-width col 1
    for col in ws1.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    # ── Hoja 2: Detalle Items ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle Items")
    headers2 = ['Factura', 'SKU', 'Descripcion', 'Cantidad', 'P. Unitario', 'Descuento %', 'Subtotal']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin
    row = 2
    for f in facturas:
        for it in items_map.get(f['numero'], []):
            vals2 = [
                f['numero'], it.get('sku'), it.get('descripcion'),
                it.get('cantidad'), it.get('precio_unitario'),
                it.get('descuento_pct'), it.get('subtotal')
            ]
            for col, v in enumerate(vals2, 1):
                cell = ws2.cell(row, col, v)
                cell.border = thin
                if isinstance(v, (int, float)):
                    cell.number_format = '#,##0'
            row += 1
    for col in ws2.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # ── Hoja 3: Pagos ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Pagos Recibidos")
    headers3 = ['Factura', 'Fecha', 'Monto', 'Medio', 'Referencia', 'Registrado Por']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(1, col, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin
    row = 3
    for f in facturas:
        for p in pagos_map.get(f['numero'], []):
            vals3 = [
                f['numero'], p.get('fecha'), p.get('monto'),
                p.get('medio'), p.get('referencia'), p.get('registrado_por')
            ]
            for col, v in enumerate(vals3, 1):
                cell = ws3.cell(row, col, v)
                cell.border = thin
                if isinstance(v, (int, float)):
                    cell.number_format = '#,##0'
            row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── RUTAS ────────────────────────────────────────────────────────────────────

@bp.route('/contabilidad')
def contabilidad_index():
    from templates_py.contabilidad_html import CONTABILIDAD_HTML
    return CONTABILIDAD_HTML

@bp.route('/api/contabilidad/login', methods=['POST'])
def cont_login():
    from config import COMPRAS_USERS as ALL_PASSES
    data = request.get_json() or {}
    u = data.get('usuario', '').strip().lower()
    p = data.get('password', '')
    if u in CONT_USERS and ALL_PASSES.get(u) == p:
        session['cont_user'] = u
        return jsonify({'ok': True, 'usuario': u})
    return jsonify({'error': 'Credenciales incorrectas'}), 401

@bp.route('/api/contabilidad/logout', methods=['POST'])
def cont_logout():
    session.pop('cont_user', None)
    return jsonify({'ok': True})

@bp.route('/api/contabilidad/me')
def cont_me():
    u = session.get('cont_user', '') or session.get('compras_user', '')
    return jsonify({'autenticado': u in CONT_USERS, 'usuario': u})


# ── Facturas ──────────────────────────────────────────────────────────────────

@bp.route('/api/contabilidad/facturas')
def cont_facturas_list():
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    estado = request.args.get('estado', '')
    empresa = request.args.get('empresa', '')
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    q = "SELECT * FROM facturas WHERE 1=1"
    params = []
    if estado:
        q += " AND estado=?"; params.append(estado)
    if empresa:
        q += " AND empresa=?"; params.append(empresa)
    if fecha_desde:
        q += " AND fecha_emision>=?"; params.append(fecha_desde)
    if fecha_hasta:
        q += " AND fecha_emision<=?"; params.append(fecha_hasta)
    q += " ORDER BY fecha_creacion DESC LIMIT 200"

    rows = [dict(r) for r in conn.execute(q, params).fetchall()]

    # Calcular monto pagado por factura
    for f in rows:
        pago = conn.execute(
            "SELECT COALESCE(SUM(monto),0) FROM facturas_pagos WHERE numero_factura=?",
            (f['numero'],)
        ).fetchone()[0]
        f['monto_pagado'] = pago
        f['saldo'] = f['total'] - pago

    return jsonify(rows)


@bp.route('/api/contabilidad/facturas/generar', methods=['POST'])
def cont_facturas_generar():
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json() or {}

    numero_pedido = data.get('numero_pedido', '')
    empresa = data.get('empresa', 'ANIMUS')
    iva_pct = float(data.get('iva_pct', 0))
    notas = data.get('notas', '')
    fecha_venc = data.get('fecha_vencimiento', '')
    items_manual = data.get('items', [])  # si no viene de pedido

    conn = get_db()

    # Resolver cliente e items desde pedido (si se provee)
    cliente_id, cliente_nombre, cliente_nit = None, '', ''
    items_src = items_manual

    if numero_pedido:
        ped = conn.execute("SELECT * FROM pedidos WHERE numero=?", (numero_pedido,)).fetchone()
        if not ped:
            return jsonify({'error': f'Pedido {numero_pedido} no encontrado'}), 404
        ped = dict(ped)
        cli = conn.execute("SELECT * FROM clientes WHERE id=?", (ped['cliente_id'],)).fetchone()
        if cli:
            cli = dict(cli)
            cliente_id = cli['id']
            cliente_nombre = cli['nombre']
            cliente_nit = cli.get('nit', '')

        items_src = [dict(r) for r in conn.execute(
            "SELECT sku, descripcion, cantidad, precio_unitario, subtotal FROM pedidos_items WHERE numero_pedido=?",
            (numero_pedido,)
        ).fetchall()]

    # Calcular totales
    subtotal = sum(it.get('subtotal', 0) or (it.get('cantidad', 0) * it.get('precio_unitario', 0)) for it in items_src)
    descuento = float(data.get('descuento', 0))
    base_iva = subtotal - descuento
    iva_valor = round(base_iva * iva_pct / 100, 2)
    total = base_iva + iva_valor

    numero = _next_numero(conn, empresa)
    hoy = date.today().isoformat()

    conn.execute("""
        INSERT INTO facturas
        (numero, tipo, numero_pedido, cliente_id, cliente_nombre, cliente_nit,
         empresa, fecha_emision, fecha_vencimiento, subtotal, descuento,
         iva_pct, iva_valor, total, estado, notas, creado_por)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (numero, 'Factura', numero_pedido, cliente_id, cliente_nombre, cliente_nit,
          empresa, hoy, fecha_venc, subtotal, descuento, iva_pct, iva_valor, total,
          'Emitida', notas, _user()))

    for it in items_src:
        sub = it.get('subtotal', 0) or it.get('cantidad', 0) * it.get('precio_unitario', 0)
        conn.execute("""
            INSERT INTO facturas_items
            (numero_factura, sku, descripcion, cantidad, precio_unitario, descuento_pct, subtotal)
            VALUES(?,?,?,?,?,?,?)
        """, (numero, it.get('sku',''), it.get('descripcion',''),
              it.get('cantidad', 0), it.get('precio_unitario', 0),
              it.get('descuento_pct', 0), sub))
    conn.commit()

    return jsonify({'ok': True, 'numero': numero, 'total': total})


@bp.route('/api/contabilidad/facturas/<numero>/pdf')
def cont_factura_pdf(numero):
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    f = conn.execute("SELECT * FROM facturas WHERE numero=?", (numero,)).fetchone()
    if not f:
        return jsonify({'error': 'Factura no encontrada'}), 404
    f = dict(f)
    items = [dict(r) for r in conn.execute(
        "SELECT * FROM facturas_items WHERE numero_factura=?", (numero,)
    ).fetchall()]

    pdf_bytes = _generar_pdf_bytes(f, items)
    fname = f"{numero}.pdf"
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )


@bp.route('/api/contabilidad/facturas/<numero>/pago', methods=['POST'])
def cont_factura_pago(numero):
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json() or {}
    monto = float(data.get('monto', 0))
    if monto <= 0:
        return jsonify({'error': 'Monto debe ser mayor a 0'}), 400

    conn = get_db()
    f = conn.execute("SELECT * FROM facturas WHERE numero=?", (numero,)).fetchone()
    if not f:
        return jsonify({'error': 'Factura no encontrada'}), 404

    conn.execute("""
        INSERT INTO facturas_pagos(numero_factura, fecha, monto, medio, referencia, registrado_por)
        VALUES(?,?,?,?,?,?)
    """, (numero, date.today().isoformat(), monto,
          data.get('medio', 'Transferencia'),
          data.get('referencia', ''), _user()))

    # Actualizar estado si está totalmente pagada
    total_pagado = conn.execute(
        "SELECT COALESCE(SUM(monto),0) FROM facturas_pagos WHERE numero_factura=?", (numero,)
    ).fetchone()[0]
    if total_pagado >= dict(f)['total']:
        conn.execute("UPDATE facturas SET estado='Pagada' WHERE numero=?", (numero,))
    conn.commit()

    return jsonify({'ok': True, 'total_pagado': total_pagado})


@bp.route('/api/contabilidad/facturas/<numero>/anular', methods=['PATCH'])
def cont_factura_anular(numero):
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    data = request.get_json() or {}
    motivo = data.get('motivo', 'Sin motivo')
    conn = get_db()
    conn.execute(
        "UPDATE facturas SET estado='Anulada', notas=notas||' | ANULADA: '||? WHERE numero=?",
        (motivo, numero)
    )
    conn.commit()
    return jsonify({'ok': True})


# ── Export Siigo ──────────────────────────────────────────────────────────────

@bp.route('/api/contabilidad/export/siigo')
def cont_export_siigo():
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    desde = request.args.get('desde', date.today().replace(day=1).isoformat())
    hasta = request.args.get('hasta', date.today().isoformat())
    empresa = request.args.get('empresa', '')

    q = "SELECT * FROM facturas WHERE fecha_emision>=? AND fecha_emision<=? AND estado!='Anulada'"
    params = [desde, hasta]
    if empresa:
        q += " AND empresa=?"; params.append(empresa)
    q += " ORDER BY fecha_emision"
    facturas = [dict(r) for r in conn.execute(q, params).fetchall()]

    items_map = {}
    pagos_map = {}
    for f in facturas:
        items_map[f['numero']] = [dict(r) for r in conn.execute(
            "SELECT * FROM facturas_items WHERE numero_factura=?", (f['numero'],)
        ).fetchall()]
        pagos_map[f['numero']] = [dict(r) for r in conn.execute(
            "SELECT * FROM facturas_pagos WHERE numero_factura=?", (f['numero'],)
        ).fetchall()]

    excel = _generar_excel_siigo(facturas, items_map, pagos_map)
    fname = f"Siigo_Ventas_{desde}_{hasta}.xlsx"
    return Response(
        excel,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )


# ── Nómina (read desde RRHH) ──────────────────────────────────────────────────

@bp.route('/api/contabilidad/nomina')
def cont_nomina():
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    periodo = request.args.get('periodo', '')
    q = """
        SELECT nr.*, e.nombre || ' ' || e.apellido as nombre_completo,
               e.cargo, e.area, e.empresa as empresa_emp
        FROM nomina_registros nr
        JOIN empleados e ON nr.empleado_id = e.id
    """
    params = []
    if periodo:
        q += " WHERE nr.periodo=?"; params.append(periodo)
    q += " ORDER BY nr.periodo DESC, nombre_completo"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]

    # Totales por periodo
    totales = {}
    for r in rows:
        p = r['periodo']
        if p not in totales:
            totales[p] = {'bruto': 0, 'neto': 0, 'parafiscales': 0, 'empleados': 0}
        totales[p]['bruto'] += r.get('salario_base', 0)
        totales[p]['neto'] += r.get('salario_neto', 0)
        totales[p]['parafiscales'] += r.get('descuento_salud', 0) + r.get('descuento_pension', 0)
        totales[p]['empleados'] += 1

    periodos_disponibles = [dict(r) for r in conn.execute(
        "SELECT DISTINCT periodo FROM nomina_registros ORDER BY periodo DESC LIMIT 24"
    ).fetchall()]

    return jsonify({'registros': rows, 'totales': totales, 'periodos': periodos_disponibles})


# ── Tesorería (read desde financiero) ────────────────────────────────────────

@bp.route('/api/contabilidad/tesoreria')
def cont_tesoreria():
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    desde = request.args.get('desde', (date.today() - timedelta(days=30)).isoformat())
    hasta = request.args.get('hasta', date.today().isoformat())

    egresos = [dict(r) for r in conn.execute(
        "SELECT * FROM flujo_egresos WHERE fecha>=? AND fecha<=? ORDER BY fecha DESC",
        (desde, hasta)
    ).fetchall()]
    ingresos = [dict(r) for r in conn.execute(
        "SELECT * FROM flujo_ingresos WHERE fecha>=? AND fecha<=? ORDER BY fecha DESC",
        (desde, hasta)
    ).fetchall()]

    total_egr = sum(r.get('monto', 0) for r in egresos)
    total_ing = sum(r.get('monto', 0) for r in ingresos)

    return jsonify({
        'egresos': egresos,
        'ingresos': ingresos,
        'total_egresos': total_egr,
        'total_ingresos': total_ing,
        'flujo_neto': total_ing - total_egr,
        'periodo': {'desde': desde, 'hasta': hasta}
    })


# ── KPIs resumen ──────────────────────────────────────────────────────────────

@bp.route('/api/contabilidad/kpis')
def cont_kpis():
    if not _auth():
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    hoy = date.today().isoformat()
    mes_inicio = date.today().replace(day=1).isoformat()

    # Facturado este mes
    fact_mes = conn.execute(
        "SELECT COALESCE(SUM(total),0) FROM facturas WHERE fecha_emision>=? AND estado!='Anulada'",
        (mes_inicio,)
    ).fetchone()[0]

    # Cartera pendiente (emitidas no pagadas)
    cartera = conn.execute(
        "SELECT COALESCE(SUM(total),0) FROM facturas WHERE estado='Emitida'"
    ).fetchone()[0]
    cartera_vencida = conn.execute(
        "SELECT COALESCE(SUM(total),0) FROM facturas WHERE estado='Emitida' AND fecha_vencimiento!='' AND fecha_vencimiento<?",
        (hoy,)
    ).fetchone()[0]

    # Facturas emitidas este mes
    n_fact_mes = conn.execute(
        "SELECT COUNT(*) FROM facturas WHERE fecha_emision>=? AND estado!='Anulada'",
        (mes_inicio,)
    ).fetchone()[0]

    return jsonify({
        'facturado_mes': fact_mes,
        'cartera_total': cartera,
        'cartera_vencida': cartera_vencida,
        'facturas_mes': n_fact_mes,
        'mes': mes_inicio[:7]
    })
