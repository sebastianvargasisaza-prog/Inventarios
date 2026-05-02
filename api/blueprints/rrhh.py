# blueprints/rrhh.py — extraído de index.py (Fase C)
import os
import json
import sqlite3
import hmac
import time
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request, Response, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from config import DB_PATH, COMPRAS_USERS, ADMIN_USERS, CONTADORA_USERS, RRHH_USERS
from database import get_db
from auth import _client_ip, _is_locked, _record_failure, _clear_attempts, _log_sec, sin_acceso_html
from templates_py.rrhh_html import RRHH_HTML
from templates_py.compromisos_html import COMPROMISOS_HTML
from templates_py.home_html import HOME_HTML
from templates_py.hub_html import HUB_HTML
from templates_py.clientes_html import CLIENTES_HTML
from templates_py.calidad_html import CALIDAD_HTML
from templates_py.gerencia_html import GERENCIA_HTML
from templates_py.financiero_html import FINANCIERO_HTML
from templates_py.login_html import LOGIN_HTML
from templates_py.compras_html import COMPRAS_HTML
from templates_py.recepcion_html import RECEPCION_HTML
from templates_py.salida_html import SALIDA_HTML
from templates_py.solicitudes_html import SOLICITUDES_HTML
from templates_py.dashboard_html import DASHBOARD_HTML

bp = Blueprint('rrhh', __name__)


_RRHH_AUTHORIZED = lambda: set(RRHH_USERS) | set(ADMIN_USERS) | set(CONTADORA_USERS)


@bp.before_request
def _rrhh_gate():
    """Audit zero-error 2-may-2026 · gate PII para TODOS los endpoints rrhh.

    Antes 23 endpoints `/api/rrhh/*` solo verificaban `session.get("compras_user")`,
    permitiendo a CUALQUIER user logueado leer cédulas, salarios, cuentas
    bancarias de toda la planta. Violación Habeas Data Ley 1581/2012.

    Aplicado a TODOS los endpoints del blueprint vía before_request.
    Excepción: la página HTML /rrhh tiene su propio gate con sin_acceso_html().
    """
    # /rrhh (HTML) ya tiene su propio gate con sin_acceso_html() — skip
    p = request.path
    if p == '/rrhh':
        return None
    # Solo gating endpoints API
    if not p.startswith('/api/rrhh'):
        return None
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    u = session.get('compras_user', '')
    if u not in _RRHH_AUTHORIZED():
        return jsonify({'error': 'Acceso restringido a RRHH/Admin/Contabilidad'}), 403
    # OK, sigue al endpoint
    return None


@bp.route("/rrhh")
def rrhh_panel():
    if "compras_user" not in session:
        return redirect("/login?next=/rrhh")
    u = session.get("compras_user", "")
    if u not in RRHH_USERS:
        return Response(sin_acceso_html("Recursos Humanos"), mimetype="text/html")
    usuario = u.capitalize()
    resp = Response(RRHH_HTML.replace("{usuario}", usuario), mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp

@bp.route("/api/rrhh/dashboard")
def rrhh_dashboard():
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM empleados WHERE estado='Activo'")
    headcount = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(salario_base),0) FROM empleados WHERE estado='Activo'")
    nomina_bruta = c.fetchone()[0]
    mes_actual = datetime.now().strftime("%Y-%m")
    c.execute("SELECT COALESCE(SUM(dias),0) FROM ausencias WHERE estado='Aprobada' AND fecha_inicio LIKE ?", (mes_actual+"%",))
    dias_ausentes = c.fetchone()[0]
    ausentismo_pct = round(dias_ausentes/(headcount*22)*100,1) if headcount>0 else 0
    c.execute("SELECT COUNT(*) FROM capacitaciones_empleados WHERE completado=0")
    caps_pendientes = c.fetchone()[0]
    c.execute("SELECT empresa, COUNT(*) FROM empleados WHERE estado='Activo' GROUP BY empresa ORDER BY 2 DESC")
    por_empresa = [{"empresa":r[0],"count":r[1]} for r in c.fetchall()]
    c.execute("SELECT area, COUNT(*) FROM empleados WHERE estado='Activo' GROUP BY area ORDER BY 2 DESC")
    por_area = [{"area":r[0],"count":r[1]} for r in c.fetchall()]
    alertas = []
    from datetime import date as ddate
    c.execute("SELECT id, nombre||' '||apellido, fecha_ingreso FROM empleados WHERE estado='Activo'")
    for emp in c.fetchall():
        if emp[2]:
            try:
                fi = ddate.fromisoformat(emp[2])
                if (ddate.today()-fi).days > 365:
                    c.execute("SELECT COALESCE(SUM(dias),0) FROM ausencias WHERE tipo='Vacaciones' AND estado='Aprobada' AND empleado_id=?", (emp[0],))
                    vac = c.fetchone()[0]
                    if vac < 15:
                        alertas.append({"tipo":"warn","msg":emp[1]+" tiene "+str(15-vac)+" dias de vacaciones pendientes"})
            except: pass
    c.execute("SELECT nombre||' '||apellido, fecha_fin_contrato FROM empleados WHERE tipo_contrato='Fijo' AND fecha_fin_contrato!='' AND estado='Activo'")
    for r in c.fetchall():
        if r[1]:
            try:
                fv = ddate.fromisoformat(r[1])
                d_days = (fv-ddate.today()).days
                if 0 < d_days <= 45:
                    alertas.append({"tipo":"danger","msg":"Contrato de "+r[0]+" vence en "+str(d_days)+" dias"})
            except: pass
    return jsonify({"headcount":headcount,"nomina_bruta":nomina_bruta,"ausentismo_pct":ausentismo_pct,"caps_pendientes":caps_pendientes,"por_empresa":por_empresa,"por_area":por_area,"alertas":alertas})

@bp.route("/api/rrhh/empleados", methods=["GET","POST"])
def rrhh_empleados():
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    if request.method == "POST":
        d = request.get_json(silent=True) or {}
        c.execute("SELECT COUNT(*) FROM empleados"); n = c.fetchone()[0]+1
        codigo = "EMP"+str(n).zfill(4)
        c.execute("INSERT INTO empleados (codigo,nombre,apellido,cedula,cargo,area,empresa,tipo_contrato,fecha_ingreso,estado,salario_base,eps,afp,arl,caja_compensacion,email,telefono,nivel_riesgo,observaciones,banco,numero_cuenta,tipo_cuenta) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                 (codigo,d.get("nombre",""),d.get("apellido",""),d.get("cedula",""),d.get("cargo",""),d.get("area",""),d.get("empresa","Espagiria"),d.get("tipo_contrato","Indefinido"),d.get("fecha_ingreso",""),"Activo",float(d.get("salario_base",0)),d.get("eps",""),d.get("afp",""),d.get("arl",""),d.get("caja",""),d.get("email",""),d.get("telefono",""),int(d.get("nivel_riesgo",1)),d.get("observaciones",""),d.get("banco",""),d.get("numero_cuenta",""),d.get("tipo_cuenta","")))
        conn.commit(); new_id=c.lastrowid
        return jsonify({"ok":True,"id":new_id,"codigo":codigo}),201
    c.execute("SELECT id,codigo,nombre,apellido,cargo,area,empresa,tipo_contrato,fecha_ingreso,estado,salario_base,email,telefono,eps,afp,nivel_riesgo FROM empleados ORDER BY empresa,nombre")
    rows=c.fetchall()
    return jsonify([{"id":r[0],"codigo":r[1],"nombre":r[2],"apellido":r[3],"cargo":r[4],"area":r[5],"empresa":r[6],"tipo_contrato":r[7],"fecha_ingreso":r[8],"estado":r[9],"salario_base":r[10],"email":r[11],"telefono":r[12],"eps":r[13],"afp":r[14],"nivel_riesgo":r[15]} for r in rows])

@bp.route("/api/rrhh/empleados/<int:eid>", methods=["GET","PUT"])
def rrhh_empleado_det(eid):
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    if request.method == "PUT":
        d = request.get_json(silent=True) or {}
        c.execute("UPDATE empleados SET nombre=?,apellido=?,cargo=?,area=?,empresa=?,tipo_contrato=?,salario_base=?,eps=?,afp=?,arl=?,caja_compensacion=?,email=?,telefono=?,nivel_riesgo=?,observaciones=?,estado=?,banco=?,numero_cuenta=?,tipo_cuenta=? WHERE id=?",
                 (d.get("nombre",""),d.get("apellido",""),d.get("cargo",""),d.get("area",""),d.get("empresa",""),d.get("tipo_contrato",""),float(d.get("salario_base",0)),d.get("eps",""),d.get("afp",""),d.get("arl",""),d.get("caja",""),d.get("email",""),d.get("telefono",""),int(d.get("nivel_riesgo",1)),d.get("observaciones",""),d.get("estado","Activo"),d.get("banco",""),d.get("numero_cuenta",""),d.get("tipo_cuenta",""),eid))
        conn.commit(); return jsonify({"ok":True})
    c.execute("SELECT * FROM empleados WHERE id=?", (eid,))
    r=c.fetchone()
    if not r: return jsonify({"error":"not found"}),404
    cols=[d[0] for d in c.description]
    return jsonify(dict(zip(cols,r)))

@bp.route("/api/rrhh/nomina/<periodo>")
def rrhh_nomina(periodo):
    u = session.get("compras_user", "")
    SMMLV=1423500; AUX=202000
    # Quincenal: periodo formato YYYY-MM-Q1 o YYYY-MM-Q2
    es_q2 = periodo.endswith("-Q2")
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id,nombre,apellido,cargo,salario_base,empresa,area,nivel_riesgo,banco,numero_cuenta,tipo_cuenta FROM empleados WHERE estado='Activo' ORDER BY empresa,nombre")
    emps=c.fetchall()
    c.execute("SELECT empleado_id,dias_trabajados,horas_extras,valor_horas_extras,bonificaciones,otros_descuentos FROM nomina_registros WHERE periodo=?", (periodo,))
    ex={r[0]:r for r in c.fetchall()}
    result=[]
    arl_rates={1:0.00522,2:0.01044,3:0.02436,4:0.04350,5:0.06960}
    for e in emps:
        eid,nom,ape,cargo,sal,emp,area,riesgo,banco,num_cta,tipo_cta=e
        xr=ex.get(eid)
        # Quincenal: base de pago es la mitad del salario mensual
        sal_q = round(sal / 2)
        dias=xr[1] if xr else 15; he=xr[2] if xr else 0; vhe=xr[3] if xr else 0
        bonos=xr[4] if xr else 0; otros=xr[5] if xr else 0
        # Pro-rateo por días trabajados sobre 15
        sal_prop = round(sal_q * dias / 15)
        # Aux transporte quincenal (solo si salario <= 2 SMMLV)
        aux = round(AUX / 2) if sal <= 2*SMMLV else 0
        aux_prop = round(aux * dias / 15)
        # Deducciones sobre el devengado quincenal proporcional
        desc_salud = round(sal_prop * 0.04)
        desc_pension = round(sal_prop * 0.04)
        neto = sal_prop + aux_prop + vhe + bonos - desc_salud - desc_pension - otros
        # Aportes empleador (calculados sobre salario mensual / 2)
        ap_s=round(sal_q*0.085); ap_p=round(sal_q*0.12)
        ap_arl=round(sal_q*arl_rates.get(riesgo,0.00522))
        ap_sena=round(sal_q*0.02); ap_icbf=round(sal_q*0.03); ap_caja=round(sal_q*0.04)
        ap_tot=ap_s+ap_p+ap_arl+ap_sena+ap_icbf+ap_caja
        result.append({"id":eid,"nombre":nom+" "+ape,"cargo":cargo,"empresa":emp,"area":area,
            "salario_base":sal,"salario_quincenal":sal_q,"dias_trabajados":dias,
            "aux_transporte":aux_prop,"horas_extras":he,"valor_horas_extras":vhe,
            "bonificaciones":bonos,"desc_salud":desc_salud,"desc_pension":desc_pension,
            "otros_descuentos":otros,"neto":neto,
            "banco":banco or "","numero_cuenta":num_cta or "","tipo_cuenta":tipo_cta or "",
            "aportes_empleador":{"salud":ap_s,"pension":ap_p,"arl":ap_arl,"sena":ap_sena,"icbf":ap_icbf,"caja":ap_caja,"total":ap_tot}})
    return jsonify(result)

@bp.route("/api/rrhh/nomina/guardar", methods=["POST"])
def rrhh_nomina_guardar():
    u = session.get("compras_user", "")
    d=request.get_json(silent=True) or {}
    periodo=d.get("periodo",""); registros=d.get("registros",[])
    conn = get_db(); c = conn.cursor()
    for r in registros:
        c.execute("INSERT OR REPLACE INTO nomina_registros (periodo,empleado_id,salario_base,dias_trabajados,horas_extras,valor_horas_extras,subsidio_transporte,bonificaciones,descuento_salud,descuento_pension,otros_descuentos,salario_neto,estado) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                 (periodo,r["id"],r["salario_base"],r.get("dias_trabajados",30),r.get("horas_extras",0),r.get("valor_horas_extras",0),r.get("aux_transporte",0),r.get("bonificaciones",0),r["desc_salud"],r["desc_pension"],r.get("otros_descuentos",0),r["neto"],"Generada"))
    conn.commit()
    return jsonify({"ok":True,"periodo":periodo,"registros":len(registros)})

@bp.route("/api/rrhh/ausencias", methods=["GET","POST"])
def rrhh_ausencias():
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    if request.method=="POST":
        d=request.get_json(silent=True) or {}
        c.execute("INSERT INTO ausencias (empleado_id,tipo,fecha_inicio,fecha_fin,dias,estado,observaciones) VALUES (?,?,?,?,?,'Pendiente',?)",
                 (int(d.get("empleado_id",0)),d.get("tipo","Vacaciones"),d.get("fecha_inicio",""),d.get("fecha_fin",""),int(d.get("dias",0)),d.get("observaciones","")))
        conn.commit(); return jsonify({"ok":True}),201
    c.execute("SELECT a.id,e.nombre||' '||e.apellido,a.tipo,a.fecha_inicio,a.fecha_fin,a.dias,a.estado,a.observaciones,a.aprobado_por FROM ausencias a JOIN empleados e ON a.empleado_id=e.id ORDER BY a.creado_en DESC LIMIT 200")
    rows=c.fetchall()
    return jsonify([{"id":r[0],"empleado":r[1],"tipo":r[2],"fecha_inicio":r[3],"fecha_fin":r[4],"dias":r[5],"estado":r[6],"observaciones":r[7],"aprobado_por":r[8]} for r in rows])

@bp.route("/api/rrhh/ausencias/<int:aid>", methods=["PATCH"])
def rrhh_ausencia_upd(aid):
    u = session.get("compras_user", "")
    d=request.get_json(silent=True) or {}
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE ausencias SET estado=?,aprobado_por=? WHERE id=?", (d.get("estado",""),session.get("compras_user",""),aid))
    conn.commit(); return jsonify({"ok":True})

@bp.route("/api/rrhh/capacitaciones", methods=["GET","POST"])
def rrhh_caps():
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    if request.method=="POST":
        d=request.get_json(silent=True) or {}
        c.execute("INSERT INTO capacitaciones (nombre,tipo,fecha,duracion_horas,instructor,empresa,obligatoria) VALUES (?,?,?,?,?,?,?)",
                 (d.get("nombre",""),d.get("tipo","BPM"),d.get("fecha",""),float(d.get("duracion_horas",1)),d.get("instructor",""),d.get("empresa","Espagiria"),1 if d.get("obligatoria") else 0))
        cap_id=c.lastrowid
        c.execute("SELECT id FROM empleados WHERE estado='Activo'")
        for emp in c.fetchall():
            try: c.execute("INSERT OR IGNORE INTO capacitaciones_empleados (capacitacion_id,empleado_id,completado) VALUES (?,?,0)", (cap_id,emp[0]))
            except: pass
        conn.commit(); return jsonify({"ok":True,"id":cap_id}),201
    c.execute("SELECT c.id,c.nombre,c.tipo,c.fecha,c.duracion_horas,c.instructor,c.obligatoria,COUNT(ce.id),COALESCE(SUM(ce.completado),0) FROM capacitaciones c LEFT JOIN capacitaciones_empleados ce ON c.id=ce.capacitacion_id GROUP BY c.id ORDER BY c.fecha DESC")
    rows=c.fetchall()
    return jsonify([{"id":r[0],"nombre":r[1],"tipo":r[2],"fecha":r[3],"horas":r[4],"instructor":r[5],"obligatoria":r[6],"total":r[7],"completados":r[8]} for r in rows])

@bp.route("/api/rrhh/evaluaciones", methods=["GET","POST"])
def rrhh_evals():
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    if request.method=="POST":
        d=request.get_json(silent=True) or {}
        scores=[float(d.get(k,0)) for k in ["calidad","asistencia","actitud","conocimiento","productividad"]]
        total=round(sum(scores)/5,1)
        c.execute("INSERT INTO evaluaciones (empleado_id,periodo,evaluador,puntaje_total,puntaje_calidad,puntaje_asistencia,puntaje_actitud,puntaje_conocimiento,puntaje_productividad,comentarios,estado) VALUES (?,?,?,?,?,?,?,?,?,?,'Publicada')",
                 (int(d.get("empleado_id",0)),d.get("periodo",""),session.get("compras_user",""),total,scores[0],scores[1],scores[2],scores[3],scores[4],d.get("comentarios","")))
        conn.commit(); return jsonify({"ok":True}),201
    periodo=request.args.get("periodo","")
    q="SELECT ev.id,e.nombre||' '||e.apellido,e.cargo,ev.periodo,ev.evaluador,ev.puntaje_total,ev.puntaje_calidad,ev.puntaje_asistencia,ev.puntaje_actitud,ev.puntaje_conocimiento,ev.puntaje_productividad,ev.comentarios FROM evaluaciones ev JOIN empleados e ON ev.empleado_id=e.id"
    if periodo: c.execute(q+" WHERE ev.periodo=? ORDER BY ev.puntaje_total DESC",(periodo,))
    else: c.execute(q+" ORDER BY ev.periodo DESC,ev.puntaje_total DESC LIMIT 50")
    rows=c.fetchall()
    return jsonify([{"id":r[0],"empleado":r[1],"cargo":r[2],"periodo":r[3],"evaluador":r[4],"total":r[5],"calidad":r[6],"asistencia":r[7],"actitud":r[8],"conocimiento":r[9],"productividad":r[10],"comentarios":r[11]} for r in rows])

@bp.route("/api/rrhh/sgsst", methods=["GET","POST"])
def rrhh_sgsst():
    u = session.get("compras_user", "")
    conn = get_db(); c = conn.cursor()
    if request.method=="POST":
        d=request.get_json(silent=True) or {}
        c.execute("INSERT INTO sgsst_items (categoria,descripcion,frecuencia,responsable,proximo_vencimiento,estado) VALUES (?,?,?,?,?,'Pendiente')",
                 (d.get("categoria",""),d.get("descripcion",""),d.get("frecuencia","Anual"),d.get("responsable",""),d.get("proximo_vencimiento","")))
        conn.commit(); return jsonify({"ok":True}),201
    c.execute("SELECT id,categoria,descripcion,frecuencia,ultimo_cumplimiento,proximo_vencimiento,responsable,estado FROM sgsst_items ORDER BY categoria,descripcion")
    rows=c.fetchall()
    return jsonify([{"id":r[0],"categoria":r[1],"descripcion":r[2],"frecuencia":r[3],"ultimo":r[4],"proximo":r[5],"responsable":r[6],"estado":r[7]} for r in rows])

@bp.route("/api/rrhh/sgsst/<int:sid>", methods=["PATCH"])
def rrhh_sgsst_upd(sid):
    u = session.get("compras_user", "")
    d=request.get_json(silent=True) or {}
    from datetime import date as ddate, timedelta
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT frecuencia FROM sgsst_items WHERE id=?", (sid,))
    row=c.fetchone(); hoy=ddate.today().isoformat()
    freq_days={"Mensual":30,"Trimestral":90,"Semestral":180,"Anual":365}
    prox=d.get("proximo_vencimiento","") or (ddate.today()+timedelta(days=freq_days.get(row[0] if row else "Anual",365))).isoformat()
    c.execute("UPDATE sgsst_items SET estado='Cumplido',ultimo_cumplimiento=?,proximo_vencimiento=? WHERE id=?", (hoy,prox,sid))
    conn.commit(); return jsonify({"ok":True})
@bp.route("/api/rrhh/nomina/<periodo>/export")
def rrhh_nomina_export(periodo):
    """Export nomina as xlsx."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl no disponible"}), 500
    SMMLV = 1423500; AUX = 202000
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id,nombre,apellido,cargo,salario_base,empresa,nivel_riesgo FROM empleados WHERE estado='Activo' ORDER BY empresa,nombre")
    emps = c.fetchall()
    c.execute("SELECT empleado_id,dias_trabajados,valor_horas_extras,bonificaciones,otros_descuentos FROM nomina_registros WHERE periodo=?", (periodo,))
    ex = {r[0]: r for r in c.fetchall()}
    arl_rates = {1:0.00522, 2:0.01044, 3:0.02436, 4:0.04350, 5:0.06960}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Nomina " + periodo
    thin = Side(style="thin", color="D6D3D1")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill("solid", fgColor="1C1917")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    ws.merge_cells("A1:N1")
    ws["A1"] = "NOMINA PERIODO " + periodo + " - HHA GROUP"
    ws["A1"].font = Font(bold=True, size=13); ws["A1"].alignment = Alignment(horizontal="center")
    hdrs = ["#","Cedula","Empleado","Cargo","Empresa","Dias","Sal.Base","Aux.Trans","H.Extras","Bonos","-Salud 4%","-Pension 4%","-Otros","NETO"]
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center"); cell.border = brd
    for i, w in enumerate([4,14,28,22,14,6,14,12,12,12,12,12,10,14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    total_neto = 0; total_ap = 0
    alt = PatternFill("solid", fgColor="F9F8F7")
    for idx2, e in enumerate(emps, 1):
        eid, nom, ape, cargo, sal, emp, riesgo = e
        xr = ex.get(eid)
        dias = xr[1] if xr else 30; vhe = xr[2] if xr else 0
        bonos = xr[3] if xr else 0; otros = xr[4] if xr else 0
        aux = AUX if sal <= 2*SMMLV else 0
        ds = round(sal*0.04); dp = round(sal*0.04)
        neto = sal + aux + vhe + bonos - ds - dp - otros
        ap = round(sal*(0.085+0.12+arl_rates.get(riesgo,0.00522)+0.02+0.03+0.04))
        total_neto += neto; total_ap += ap
        row = idx2 + 3
        vals = [idx2, eid, nom+" "+ape, cargo, emp, dias, sal, aux, vhe, bonos, -ds, -dp, -otros if otros else 0, neto]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if idx2 % 2 == 0: cell.fill = alt
            cell.border = brd
            if col >= 7: cell.number_format = "#,##0"
            if col == 14: cell.font = Font(bold=True, color="6D28D9")
    tr = len(emps) + 4
    ws.cell(tr, 1, "TOTAL").font = Font(bold=True)
    ws.cell(tr, 14, total_neto).font = Font(bold=True, color="6D28D9")
    ws.cell(tr, 14).number_format = "#,##0"
    ws.cell(tr+1, 1, "Aportes empleador:").font = Font(size=9, color="78716C")
    ws.cell(tr+1, 14, total_ap).number_format = "#,##0"
    ws.cell(tr+2, 1, "Costo total empresa:").font = Font(bold=True)
    ws.cell(tr+2, 14, total_neto + total_ap).font = Font(bold=True)
    ws.cell(tr+2, 14).number_format = "#,##0"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Nomina_"+periodo+".xlsx"})

@bp.route("/api/rrhh/nomina/<periodo>/comprobante/<int:eid>")
def rrhh_comprobante(periodo, eid):
    """Print-ready HTML pay stub for one employee."""
    SMMLV = 1423500; AUX = 202000
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id,nombre,apellido,cedula,cargo,empresa,salario_base,nivel_riesgo,banco,numero_cuenta,tipo_cuenta FROM empleados WHERE id=?", (eid,))
    emp = c.fetchone()
    if not emp: return "<p>No encontrado</p>", 404
    eid2, nom, ape, ced, cargo, empresa, sal, riesgo, banco, num_cta, tipo_cta = emp
    c.execute("SELECT dias_trabajados,valor_horas_extras,bonificaciones,otros_descuentos,estado,aprobado_por,aprobado_en,pagado_por,pagado_en FROM nomina_registros WHERE periodo=? AND empleado_id=?", (periodo, eid))
    nr = c.fetchone()
    dias = nr[0] if nr else 30; vhe = nr[1] if nr else 0
    bonos = nr[2] if nr else 0; otros = nr[3] if nr else 0
    estado = nr[4] if nr else "No guardada"
    ap_por = nr[5] if nr else ""; ap_en = nr[6] if nr else ""
    pag_por = nr[7] if nr else ""; pag_en = nr[8] if nr else ""
    arl_rates = {1:0.00522, 2:0.01044, 3:0.02436, 4:0.04350, 5:0.06960}
    aux = AUX if sal <= 2*SMMLV else 0
    ds = round(sal*0.04); dp = round(sal*0.04)
    neto = sal + aux + vhe + bonos - ds - dp - otros
    ap_s=round(sal*0.085); ap_p=round(sal*0.12); ap_arl=round(sal*arl_rates.get(riesgo,0.00522))
    ap_sena=round(sal*0.02); ap_icbf=round(sal*0.03); ap_caja=round(sal*0.04)
    ap_tot = ap_s+ap_p+ap_arl+ap_sena+ap_icbf+ap_caja
    def cop(v): return "${:,.0f}".format(v).replace(",",".")
    if estado == "Pagada":
        badge = "background:#166534;color:#fff"
    elif estado == "Aprobada":
        badge = "background:#dcfce7;color:#166534"
    else:
        badge = "background:#fef3c7;color:#92400e"
    aprobado_txt = (" &nbsp;|&nbsp; Aprobada por: <strong>"+ap_por+"</strong> el "+ap_en) if ap_por else ""
    pagado_txt = (" &nbsp;|&nbsp; &#128184; Pagada por: <strong>"+pag_por+"</strong> el "+pag_en) if pag_por else ""
    html = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        "<title>Comprobante "+periodo+" - "+nom+" "+ape+"</title>"
        "<style>body{font-family:Arial,sans-serif;font-size:12px;margin:20px;color:#1C1917;}"
        ".hdr{text-align:center;border-bottom:2px solid #1C1917;padding-bottom:8px;margin-bottom:12px;}"
        ".hdr h2{margin:0;font-size:15px;}.hdr p{margin:2px;font-size:11px;color:#555;}"
        "table{width:100%;border-collapse:collapse;margin-bottom:8px;}"
        "td{padding:4px 8px;}tr:nth-child(even){background:#F9F8F7;}"
        ".lbl{color:#666;}.val{text-align:right;font-weight:600;}"
        ".sec{background:#1C1917;color:#fff;padding:3px 8px;font-weight:700;font-size:11px;margin-top:8px;}"
        ".neto{font-size:18px;font-weight:700;color:#6D28D9;text-align:right;padding:8px;border-top:2px solid #6D28D9;}"
        ".footer{font-size:10px;color:#999;text-align:center;margin-top:16px;border-top:1px solid #ddd;padding-top:6px;}"
        "@media print{button{display:none;}}"
        "</style></head><body>"
        "<div class='hdr'><img src='data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAIAAAAiOjnJAAAwwklEQVR42u1dd3gVxfqesntq2kmvpBdCEggdQ0dFUJAioggXC9aLiAVsqFwsIKKIV+xwpTdRpCi9BQgdQhoBUiC9nJPkJKfuzszvj00OIaDXe396ETPv4+MT9mydffebb975vm8gYwxwcPzeQLwJODixODixODixODg4sTg4sTg4sTg4OLE4OLE4OLE4ODixODixODixODg4sTg4sTg4sTg4OLE4OLE4OLE4ODixODixODixODg4sTg4sTg4sTg4OLE4OLE4OLE4ODixODixODixODg4sTg4sTg4sTg4OLE4OLE4OLE4ODixODixODixODg4sTg4sTg4sTg4OLE4OLE4OLE4ODixODixODix2i0opYRS3g6/EQJvgt/EKsYQQs1/QMgbhFus38dWIQhP5Z0/fCYTQUi53eLE+v+DEIoQOpGdO+ilWQNfeWProSMIIZkQ3jK/DsgXG/83tgqh8uqagS+8UmKxIIR0lO2d/3ZybAyhFCP+WXKL9Z+DMcYAkJzS5HkfXKw1dvDwjPfxqXU6HnzvgxqjCSNE+TfJifXfdIKUYoSeXbR47/lLKgQXPf3Y92+9GqDR5FTXTJz7geR0Msa4vefE+s8gEyJgvHDN+m/2HQSAvTPh/rt694wICV76wrM6hHbm5U9d+E+MEKEUcGrdCHj27Nm8FW7Iqq3ph5/+fIkE2KMD+r7/zBOEUkppfHgHN4x2nT135nKJmpL+XTrLlCDubHGL9Vt6QAHjrAsXp3z8qR2AgbHRi6dPJZQiCAWMZUKeG3/f00MGUgBmr/1u/e69ykbebpxYvzoMZAwjVGMyPTh3QbVTijV4rnxtplqthgBACAEASvf38fS/352S6ADw6cVfHs3KFjAmnFucWL8yDASMSU7nxHfm59YYfUTVmtdmBPv7EUpdnR2EEAKIsLDs1RmdgwLqJPLQ3A8vl1dgjLlwyol1I1YBoBDomYX/3J1/UQ3oV9Oe6toxQSakjV6FEKSUent6rnn95WCttshsnvDu/CaLFUIuCnJiXe9aESJg/P63q5YeOAQonTN+3OiB/Z2ypMzhtAEE0CnLCZHhy1561g2hI8WXn/hgIQSAUMq5xUeFbYeB63ftnbZkGYVwysB+c595QiJEJQjwF4ARkgmJCQvz1Wq3nzqTWVYh22y3d+/aut9sz+DWu1kIPZV7fuhrs02yfHt8zE/z3oYYY4RKKiuNDWaMces2ggBQxkSEE6MjFUa++vnX72/9WQRgydQnJw4bqmzkxGrXxKKUIQSrjcb+L7xyoa4+weC176N5ft7eEIBj53KGvv4WRQgAABEEADIAAAOMUYyQucnyxvixcx5/xCnLKozHv/XO+lNnvVTC9jlv9kruxGcS2/XDM8YYYISQR9//6ILR5COKq197KcDHhxACIYQIMohErVal0VAKKKGEUEqJKKrUWi0SBQChYsAYAF/PmN4tJLhekicv+LjaaEJ8JpG1Y0iyzBh7/YtvwF2jNHeP+X7vftdGQihj7NDps24jx/ndN+Hng0dOZecdy8o5mZM3+PmZeMjdr3y5hDEmE+L6//mi4qD7HgLDRo969U1KiCQrrnw7Rfu1WIrCvuXAoQWbf4IAvD723tGDBsiyyz1ijLEwfz8BQkhBr6SOXTsl9ExK7JaY4K13I05nkK9Ps/QFgOLIx0eEf/HsUxqMNp05N3/VOgGj9qxstVNiKRHGFdU1075c4mD03q6dZz08SSYEt3K6IYR2SaIMMAiabDZKqVOWKaWEUYCQ0ym1PqGAsSyTkf37vnD3XQCC97774UR2Lsao3YbJt1NiMcYghK98vbS4ri7cYFg8fSqDEKHW4ewQACAgjCAEECAIEUIYIYSQgBBgDCHUZtyj0OitR/82ICba7HQ+/8U3TqfTZdU4sdqLvrDz6PF1GcdVCM1/dFKwny8lpDWtGGCUMbPNYpckjDDE2EURlShCxix2G7w2q0L5p0qlWjT1SYNWe+RS4eebtmKEKOXEaje2Spakt1etczile7un3j9kUJtOUNkNQbhmz0G70xnm7e1nMDDQHHkVGxrCVOKWI8dtdju6dhpHcbY6x8X+fdgdjNGFP26pMZoQgu1whIjaoblCEG46eCjj4iUfd/3syRNZi7Fp49ennzrz5Y7dAMIHB/QVRYESgiECADw0eKCvp+eJ4svzV65DqK0XpYQsvzD+vtjAgMs1tV9t3gbbZWJPuyMWQohS8sVPO4lExt3WOzEqkpJrxEzKGEKwxmR6fNHiBknqHx/79OgRSl4hQpBQGhMeNn3kcADhx9u2nz6fL2DcmlsKjQweHk8PuxMwsHx/er25UcC4vXlaqB2aq5N5+UcvXHJz0z1+zzDGGGhlrRhglDLIwNMf/TO/utZXo/7i+amiKMqEUMpkQiilkkxeemBcn6iI+rq68XPmVZvqGGOyTAihSpSp0j9OGnpHZID/xYqKbRnHlEtzYv2VHSwAwI+HMyyWpr4d47rGxzEAWpsrQqiA0XvLV208edqg0655+YWOUZEYIZUgYIwEjEVBEAWsVqs/euqx4d1Tn7znLsCAgLEgYIybh42K7O5r8Lq7eyqTpE0ZRxVL2a6aun2l2GOMCZEPZOcCCEf07AGUzMEWt12ZPP758NF/rP3O3c09yMP9cHbOzydOYYgoZQgBBoDFaru9c/LIIQN7JydtWzAXAFBrMmWcy643NzpkWatWBfv5xoaFCoLIGBvZp/cX23cfv1hQU1fnZzAo4wZOrL8alOzTovKK82UVbnq3fsmdAAAuiYFSKmBcUFL66MefqnV6NYKlDQ2zN/wAAQQAiAJ2Wq1qUfXa+LF3pvURMD6albPp0JH07Nzi6pp6q9VJqTJVDRm7vUvKlvf+ASHs3jEu3N+vuKr69IVLQ3v1oJRhzIn1F+0H86+UmsyNncLDosNCXeKTkh9otdkmzf3A5HB4anWEyiJCfl5egDFREGrM5gFJictfndEhKDD9TOa8NRv25+TZiKzT6tSioNbr9QgpNDXb7CV19RAAxpjBwyOpQ0hB8ZWsS4VDe/Vg7SlTrD0RCwAAwKWyCiZJEX6+Oo3GVTpGyaF4dtHijMLLAV6eTlmGADIAZFkWBaHCZBrbo9vaN1+VZPnRuQtWHjwEBdHg6aGntK6xSZKcfh7uNkkGADAIKSUB7m4AQpkQEeOY4GAAWH55ObhO1ODE+kuhtsEMCAn18W62YRBSSjFCC9dsWLr3gL+3t0QIBJBBABgTBWxsarqnc/KGOW9cLqu49405meXlgT7eDolUmUwGjeaBPj1TwsO2nzx9oqRcxBhCKMlysLe30rcCjKMC/AAEtQ1NnFh/dcVBcgIABFFUiKVM7xzNyn5l2Uofg6E5iws2K15NDmest2HFazPLqmvufHlWUX1DiK9vZV1dgE4/7f4xTwy/K/NSwayV6/IqKrVqFWMMQkQp7Rge5up5RZUaQCjJUntr53ZHLGUM6DIeyutfvz9dglCA0PX+lURC2eH4YtbLHu5uI1+ffamuLtDLq7y29r5ePRY+80RIgP/Ln389/4ctGq3WTa0ijCkhy2qM0zolui6BAAQQIuCiKyfWXxQqjQZAaLHZWtOrytwoINx6Rg9jZGpsGtOj24BuqfNXrT14Pj8kMLC8pvbVUSPefWpKQ2Pj4Gkv7cvL9/PxppQQRiGAEEKr09kxJLhHx3gGmisA2p0OQJkgCgAAxkD76Qzb3ZROiI83EHBhdXXr0Je4oECJXFOCgTIgAPDqhPFmi+WTzdsM3t6VRuMzQwe/+9SU8uqaQdNf3n+xINDXhxCZMQABZIBhjK1W6+QhA9UqFSFUoVBBZRVgzNfTA7Sz+Jl2RCzFPsUEB4saTXFldU19vStxfvygAW6iIBGKEFRUA4vN3jUyIjU+dtWO3RUNjYSQpJCg+U893mSxjp/9XmZFpb+Xh1OSlAwLBhiGqMlu7xQSPOWeYcoYUxH0cy+XAITiQkIAAO1KbmhHxFKUhcSIDsEGQ6mp7uzFAsYYYIBQmhAZ/v7DE00mo0MmgiCIouiUnHf37A4AWLPvoEarsdts70yepNNqX1z85aGCAj9PD6csuzo2BBFDyGGzLnpqipter1gmCGFlbW1OSamg1abGRrW3UWH7sliUMT9vQ5fIDrLdse3YcQghAEyp8/HM2FErXnrOR6WqMdXVNJipU7qrR7dqoym3tJwy1j0y/J6+fXYdO/HN7n1+BoNTUa0AA4wJCAIIjbXGhVMeGdKzO6FEOSEDYN/ps2VGY7ifT9fYWNBK5efE+qtBiYsa0asnFIVNx07W1tcrbpYS5znxrjtPf75o2fRnHhmQNiy1c6fIiJP5FxqdDklyPjCgHwDgvbXrBbUaMAYhgABihARBMNsdTY2Nnz815dn7RsuEYIQVDkEA1u5PZ4QM7NTR4OlBKOU61l/3M0IIADCqX9q76zYW1dQs3bp95sQHlLlnJdbK4OU5adjQScOGKvvnlZRKlLqrVXf27HGlsvJs8RUvdzdGKEJIItRssRHJ2Ts66v3HH+mX2lkJDwQtoc/pZzJ3Z+dqtNpJdwwB7Q/ty2IhCGVCfLw8Jw8ZAAD857btpVXVrgpEGCElSZBQKskyAMDhcDDK9BpNRFBARm5+fUWVsa7eWF9vsVg8RDy8c9LqF6elL1qgsErx1pWsOkLo7BWrrQ7HgMT4fl1SFHe+XTV1u9OxFPZMG3Pvyn0HL9XUvrj4y3Vz3iCMQcaUah9Cq7wJhDFgFGNsczh6dYz/+pUXBFGlUqkiA/2jQ4L9vb1dPayLNzIhoiB8uGrt/vMXdCrV7EkTmsOX29l6Fu2OWBBCQqnB0/P9R/52/wcfbzx15v0Vq1+eNEGSZVHAbeTxEB8fgJAkyyZzY1RI8JRRI6/x2JSUX4RcApgky6Ig7Dl24u0N31NKp94zrHdyUvus49Ae078wQoSSMYP6P3vXEELInPU/LNm8TRQEmVytbqU42gkdQt3UalOTJf9KCaHU4XTKhCh9pZLGgxXhCwAGmESIKAhHMrMmLljU4JQHd4yfM+VhJRia+1jtx9lChNL5Tz8xKrWLVZaf/frbhWs2CBhDCJVKtQgiAEBSVGSEr68kk81HjjZnq2IsYIwRaj3EI5QCAEWMt6UfGfPO+5VWa3JQwKrXZqhVKqV7bY9fb/ssvKa8bYTxvWm9z+Wdz6mo3JuTV1R8pV9yol6nYwBQSgilapWqpKLiUP7FwurqYV1Sgv38nJKslDSCEDDGKGUAAowQkaT3lq96fumyOpujc3DglnfeDAkIoO24CFu7ro+lBPo5Hc4nP/x4eXoGhbBTgN8r48c+ePtgLAiKKSqtrOrx7ItGSUry99sw6+W4iHDQ3Pdddcd2Zhx/b82Gg5cuMsqGJXX69pUX/X192nmJrPZeeM2V4PDZdz/MWfNdVVOTIOA+MVGP3D54WO+egX6+AIBdx07cP+9DG2W+atWzI4aP7pcWERyoUqkqa2v3njqzet+BPTl5dqdk0Ommjxg+a/JDCGPa7gtG8lKRzQHvCML84svvrlz748lTZosVINTB16dHdFTXmOiUqMhTuXmf7twrQdDY1GTQaiP8fX3c3S9UVBbX1ABC9Vrt0M7Jrz80vmvHBNaSnt/OW5UT66oDrvRcJ7Jzl27ftTMzq7CmBtgdAACAscHgBREiStcpyxabDTicQBTCfX1vT+708NDb+6Z2Bi0FvfkCrJxYbV0uwIBCjFpT3eHsnIycvJwrV0qMdQ1NTU5JZhBoBEGjVof5+nSOjLitU2KfTh0DfH1ajmW8XjIn1i/Ti9I26dF2h8NqtVkdDgaATq3WaTRaraa1qYPtL9GZE+u/d7woZQywNpKVy7Ypywgg1E5lKk6s34dkoCUtUSERJxMnFsfNAfcMODixOG4d3LSwGSWJ73ohUZEr4bV+DGUM3sizueHOV391hVX9Vy4RvbqUF4QIohu58G0di3/nfjHGlGAbZc8bjiV/5aGatVwAXdNJ1y93Dv8cLiD3sW6AG+qcv1Gp+qXdFD61mT1kALA/YPLnzyCq3RxiyYRUGY2MMn8fb5Uotv7J3GQxNzaq1Wo/b4Py4dodjlpTnSiK/j7ebb5FY129zW5zc3Pzcndv88IAAJIkVRtNAIAgPz+M0W83VMoryS0sKiwpa3TYPXT6xIjwyNDgax5BliuNptYjRMiYWqXyMXiBlunt689ZVlWdXVhU29DgrtFEhYQkxUaDVqK/8rx1ZrPFYtXpdd4eHm1urMpoIoT4eHlqNZrmLbVGwiiESMmaBYypRNHX2wBa1p9qL12h0uKVNbV9n3upzmLbP/+d1MQEpd2VpIYvNm1+a8myId1Tt34wV6kEtP9M5uhZc7pERRz854eiWq1MGysv4+8ff/rj/oMvTBj/7pOPtV7MjVAqILRg9fq5G76njC2c8vATo0f+ltXelDv56VDGgu9+OFlU3GhuBJQCAfsZvPondpw2ZmRachIAEGNUUFqWNn0mUomSJNOWREIvrSYlPPytSQ92S0y4WiOJUoRQYWnZ28tW7czMKjfVAacTYOTm7pYWF/v6hPH9unZRHkdZjPOtpcu//n7zo/fevfjF55TtyiObmyyDX3qtuKJi4xuvDu93GwCgqcnS//mZJruDUkV0AxACD40muUPozPH3DeiWSm/erOXN8bEgAAQCB2NKbX12LfPsDEitNsmM2QGQrqpIV/+SALCztlXUlR6n0WL5ds8BolLJhHy1fdcj9wwXBMx+tTKH8hbfX7Zq1ur1MqP+Xp5pPVINen1RtTGrtHTj3v0FZWUZixeJqmYT6wQAMxDk4y1izAAElBTX1G45eeZQbt6uuXO6JiZQpcwfQofOZj7w7oKyhnqNWt0nPjYywM/YZDlVWLQjM/tAXv4njz/y+KgRSrQgAECizA4Aua4jYRDIANgZJJC5DLMEgIOycG9vUcAMAMBYuanup7NZe87lbnrrlbv69LpZ0TvC/55SSpsgAF1OaOuXjRWv9loOQQAwhNeWN4YAAMQYRKCNvVfSsH5MP3KpoiIuJIhCkFl8edeJk8Nv6/0rRkt5ARt273tl5Vq1VjOhV/c3J02I7hAGACCyfDIvf+6KVS9OGK/RqGWZKP6zgBCTnGtmTE+OjZEJgQDkFBQ9/c/Pj14smLNy7Y/vzSaMYQRLKqsefG9BmdXSJSJ8wZSHB6R2VooonS8sem3p8i2ZWc988U14YMCdvXs6JEnACDU/7w2aDkEA4dUPEQKAEZKc9mXPT+2amEAIgQBcrqx66cslm06eeXP56iHduwrCzbEdN8e/YwBQwBgEhFIlitz1H7mR06fYNdJ6NyLLhDAA2HUVEZQc1BW79zKn8+EhAx8eNEC22pbu3AV+eUZPyYmw2R3vrtsABGFkaudlr78c3SGMMiYRAhDqldxp0/z3+nXp7PKWWPOoE0CIEEKiICCMuyTEzRg3CkKYVVJqajBjjCGE81atK61viPH2/nH2rCE9u2NRVHquhKjIDbNfHxIfJ0M4e+VqSZJcq2MoZ75hw7G2X6IycY4QQlgQkCBEh4XOmviAVhQKqmsKyytu1vIF6H9PqatfIGMGvV7AWCWKAsYalUrAWK9W37BRMUQarVYpiO06RCUI13akzZXcj2Zlp5+/qPf0HDd44NgB/XReHrszc/IKC9EvtLIiZ5zMyztfUeWpUb31twkMAJkQBKGIceuuBCGkXBG2+U5a7pkxoKRKM8AgAObGpt3nsgBjz40c3iE40ClJyow1glCSZSwIsyaO1wpC5uXSM/kXhBbKNlvp69WTa1uRNbv7jChLdFKqPJ2ylDVljJCbVlz+5qV/USqqVPM2bFTqVCuvHCF09lKhoNU0V9ZreVUIoWqr9ZWvliCElcGTcsj58gqoUl2vJy3ZvstmsYzo0yMqOIgx1i8hfseJ08t27pn3VBS90RqNipUsKCt32GwpHeLiwjuAlgCHY1k51aY6QRQoow6HMyY0JDk2puW1MgAAAgwygBiTCTlXVLzgu00MwbjAAB8PDwBAYUVFVb3ZXaft3yWFMda6I1aWq+gSGxvh65NXWpZ7paRnUicXb37LYB3CZm9Lp1a5ctBqTaa5q9bZJDnazzcqJIixmyNr3UyBVBCFpT/vAg4HgBA0i4IMG7xEtYYw2trWQ4xqbfb3l60GlLX2MvQBAUhUuZbXopQpK4RvO3kGq9VPDB+mJMlMGXbnrsysjUeOvTJhvKeHxy/VW5dkGTDmoVGLgqAYIQzhjM+/Tj9xCrjpBY1abmyaNnrkohefc6mvSKV6eMEirahiENicUlFtbZ3FosPC6w/er9yo1Sk5KNGrVDq1GkLYhi4QQrUoaFUioMzqkFosOQQMwBvWPIJtvgcAKBPV2mlfLnHTaAADDlm6UF5ZVFMLKH39gXEatbp9Oe8QQISQ0+F8f8rk8MBAxhhEkBKGEdx29PiqA+nCtQtSUkJD3Nw+n/0aRs0Lyiu64qdbfjqUlYNbvHfKKAJ45a49laa65MjwxPDQ8qpqCEFKRHhMSPCF0rKN+9MfG3m33FJkoQ0MHh5QFK8YjQ2Nje5ubsrGe3r3DPb302rUmZdLzhVf1ul01z4PyqmoopRIhGo0GhVG93bvNmPc6LQuKYQQjHGQwctDq6lvslypqo4KC1Xu0EULxmhVXX1FvRmLYgdf76t0gYCwtoXaGAPNg2jWyplgDGF0KC9fkiQl+cig0/WJjpx5/9hRA/vRm5fQcZMsFmQAQiLLo9L6uPJeFFTU1izftae1lw0BYJR6qtUPXldd44dDR5gsucwPxthqs6/cm673cC+pq+82bYbiPGGEGMYane5fu/Y8fPdd17e1kkXYJTbGx82tuMa08/iJcUMGO2QZAjDzbxOUfaZ+9MmZ7FzXcBUCCCGiknP9zOcTIjpcrqic+MHHtVZralR4WpcUpyyrBIExFhYYEBMQcKT+wurd+wf36EYYay47w5hEiFoUNx5Ir2owB3p6dEuIV4y2SiUAjCvq6pQvCkJGKYUQWm02s9UmCIJOo7760WHodNjXvjQtMSJCphQhpNNowoODmiXDmye+3yznHULAIAANTU0yIU5JkgmxO50yIVaHE0B0fT9AGLXb7DIhkiy7DpEIcXUPSnXGrYcz8isq9CoxxNMj2NsQ6uMd4uMdaPAK8nB302hOFhSnnz2n6KvXDiQhoTQmLHRYameJkFnL1+QWFqlFsfWLsTocWBBa9dEMAgYojQsLiQ0Lvb1n99kPjqOUvb1h0w/7DqoEgVBKGRMwfnLYnQDC9cdOLN28TRQEjJCyXqtaFDMysz/8/kdKybi03kF+vk5JAgD0jIuFKlVG/qXT5/NFASMIlRTZtXv2mSxNfm76pMiIq04CgIzQuLAO8VGRnWKiO0ZFKqy66RnYN83HUoopYoQFjAmEihURMBYQwghB1FrHghghhJGAkeLwQgiVQzBCSlYyAABhBABYsXs/lcldKUlLZjzvkgYUuzV+ztxNR48v37V3YLfUG/bRjLG3H5ucnpt3obrm7jfefmb40EGpKW5a7eWq6h+PHN2RlUcovXpfylK/ADicTmVhsKfG3Lv/XNa6Y6dmLPm2T1JioJ8vpZRSNumuO3aePL0q/fC0JctPXrj04OABAT4+jVbLT0dPfPbzjsp6c9fw8LcenkgZEzFmjI3o26fT2g3ZVdWTP/h41oPjusbFSrK89cixDzf/RBzSmDt6Bvr5ugQ5jCDGyOZwKKNCjBCAELa0Z7sjFmPMbLU2WazXL7bmkGTSZLHa7Vd9akqI1dJktV1/HovdQZosdqcTACBifPBM5tZTpwGEj911pyCKrReeQBBOvnPID4cz1h48PHP82ITIiDapfwghSll4UODWOW9Mnr/wVFHxzG/+JWg0akGw2B3A6YQa9QOD+j13/1jl/VHKzBYrYZRQhhBEFDLGPnn2mdMFMy6WlE2au2DL3H+oVCpFy13y8gs6tWrp/vTPt/78xa49Oo3GKUmS1QZUqtuTEv818wWDh4fSc1FK3fX6f7303Ni352VfLnlg3kduOq1Mqd1qAwgOTe383pRHXM/FGDBbbTaL1dXr/XlC729Cij2E0O505l0sCPPxHtXvNl+Dl2vkjCAsq642NjT0iI+9s1dPZUu92VxUUpoSGTE8rQ/GGMBmwR5CmF9YCBkb2LVLt44JAIAtBw/bGs1Du3d5cvTI5lh0CEHLgoThgYEllZXuAg7xMSTHxFw/j6ZoiQE+PpNuHxTi6Qkg0olYq1bFBQcO79lt3sOTZk580E2nA4xBCG0OR35hUbSf38j+aT6eHgBCxpi7XpcUHlZpNBLJGebjG9MhVLGvgiCM6HvbbXHRCEIRIZUohHl53ZbUcdb4++Y/PcXT3c01YQwhpJSFBPiP63ubAABmTMCCv7s+NTpy+r33fPj3J3U6rbKbUmYi+8KlYC/P0f3TfA1e7M8UM83DZm48Td5sPu12hyy767QQXS2a9Ssvr7WQ0XqczxhgoPm0siRZHQ6NSlSp1Ncf1XouXPmtyWrFCGm12l/a+c+Jm0YsV/7xDbf/14F+9Jcj+xTl6bfEwSlLoSj6uIsl4NqcsF+6VnOkw42u0uYkSkDCLyW4KvfgkkUYYwrb2pz2l5qRW6w/Nf6tifovx8X/yWn/iHvgxOK4VSH80V+8kliMIFTUo9ax3opUiCBkAFBCMUaKEgggVCJAMMZKj6BM3EIIlTlE5QwtWaNXf1L2d1WSdZ3f5fEoaagYo1Y3hhCCrEW3vGZ9aEqV6AWlbrsyEqSMKjemTPeilp9a3ydoCZcAABBC2vRfhFAAWk5LKMIItJScbN0JKp0ma9ZvYcsNNNf0Ji1Xd8USAgCVTrVlAbPmZ1EiKSCAvz2G9lb3sdqWe/2Pwh1v6MO2juO7Pqbvl9xe9v9Yl+umu9K/cgM3tw/9o+QGyhiEMPvipa83bT1wLqu8sirQx2fR+o37jp3UaDWh/n4Qgu/3p9udUqC3d1Vd3Q/707vExmzPOK5SiSJGn67//sCp06nxsWW1xs82bso4c87X093Hy+vL7zf/sHufSqUKDwo8kXv+601bjmbnRoeGqETxi01beyUm1Dc2rd21p0tcLATg+wPpdqcU6ONtsztWbN/ZNT6uvrFxxfZdXeNjrXb7onXf/Zx+xM/HO8DbcLmiYvHa7y6UlCXHRCEEAQMQwvW7963e/FOD3dExPPz7fQfiwsMPZ2at2rYjIyena3zsgTPnaurrQ/391uzamxwdaTKbF67ZsPPI8ajgwPpG88mcvMiwUMDYl5u2xoSFalTKaoZQkuUvNm3dl3H8ckVlcmzM0i0/d42PzbpUeL74SkRQoEwIQqi81vjRynUHTp3pGBF+7lJBldEU7Oe7fs/+hPAO6WfOaFSquibLJ+s2HsnM8jd4+Xp5AQC2ZxyvbTCH+vs5nNLnP2zZeeiIzSnFhoVCCPecOL1k4w9FldWp8bH/S579UeYRAkApDQ3wD/L3PZGb1zsp8XxR8ens3DGD+r3++ddWm93hdM75dsXSLT9BCCx2x+T5C49kZR/NyTOazRt27z+Wd75HYiKE8GTehYqamqG9uocE+C9YtfZiaemIAX3fXbayqKLiUOY5NQSxHUJf/2KJKAjPf/7NN1t+enfF6o/WbwQAWO32t79d8dWPWxUl8dUvl372/Y9vfbPs/bUbIYSvfvYVxsLtvXq++MkXFrv9ozXfOShLiOig9JUQwk0HDm1Ozxg5eACjFECwfs9eCMGOo8c1KrFvcidREHafzbx/9rtlNbXfHTzMGHvl0y/9Dd59uyRZHY6Ll6/sPJIBGDtzPv/1r5fuPn7S1XmZLdYthzPSUjolx0RBAP6xYvX8VWsvlJX/dOKkYmMkWX7pk88SIyNS42MtNtvhs5nHc3IopSt37pEp2X3idJWpLrOgKL+0PDU2ZvqixRVGE2Bs7so1n2/aDACoa2o8cPLUgG6pX/y45VBmdl5x8Ydr1o0c2D+7oPCjNRuun8u6BYkFIQDAy8MjtWN8XFhYZGiIVqNplJxfbv45rXOKTqtZu2f/4C6dG5qaKowmjNGkwQOWbN5WWF4BARwzZGBUcNDWwxkWu8PTw624vOpYbp67Xn/ifP4rDz/Uu3PyoB49jubmhwcH7T6btXLnvlH90uotlsl3DNxy6EhJeeWwrp0BAN8fODSoc0qT1XqlsooB9tCgAZsOHi43mu7p2d3mdNaY6l54cNzA7qnJMRFZBYUzJj5QU1d3PCeP0OYI1uO55x8ZOaxXStLYwQMAAAYPTwiAh7tHSXlFtbFOEAQ/T88BKZ1eXvxleGCAxWa3WC1T7h0+9LbeiVGRhFIvNzcI4Vdbfppx/9jNhzJcmrggCCqV6kxOrrtWKxMyqk+vi5dL1u3dHxEUCAAQBaG6rkGmbNwdg0YO6BcVGkIp9TMYEEKe7m4QQg83vagSMIKdoiLvSusdHhxSWFF56Fx2YlgoYOxCSam7Vhse4N+/W+rUsaOyLhUcysoZ2T+tV0rS8w/dfzg373/ZFf6xDh0DoN7cWFNfzxgzWyzh/n4JkeGS02lzONbt2a8VBbvDseLnnbJMvNzdp44bs2LHLq1aVXClZETabbUNDemnz8qy3C2p46OjRjDGenXs+M6SFQdOntl7/Hi/5E6FpeUj+vQK8/dlkFHK6pqs3746Y/4zT1TX1dud0ro9+zWiIMvS0m3bBYwbrJaVb74y76kpRVWVGlEM9PWZt3zVjozj+ZdLooKDsi5eevSeuzYfyrh4pVSRjtJSkr7ZvO3gqbNff/8jpbTObGaA1ZsbvP39NHpdo9VaWl0zedhdof5+O06c1Ou0Wr3+8+83f7d73+m88xghSZYrjKaT+ZfsdsfF8vKdx042q6OybLHZouJiqxubJFmuMze9/eSju0+eqWtsAgBIshzg7aUW8LKfdqz6eWduYVFSbMz2E6cOnDxd32AWMTaZGyRZZgCcOJ+/Ye+BKxUVyZERX235Sa9WQ0q/3vKTShTzS8u3Hcr4bOOmlNjo/inJm/YfOnj67LxlqwZ3SW5RO279KR0IIWPMU6dLjIqglHnpdZOGDy0qL3fX64K9vac9cN/daX1KK6sSIsMho0O6dwsP9E+MDJdksnbH7oSo8AlD73BKUpCPISokmDJ2W0pSeW3tkbPnnho7qlNUpMVu6xoXPXbwwIysnK5xsVSSe6d00mjUEAJfLy+dqHr+ofEj+t5WXl2TEBlOCOmV1Emv1RBZ7hwb079L55P5F3ILimZMfCAiKKiorPzng4fvHdjPNT8d1yEMQ7g743hyXExceJgsy0nR0QCwsxcLLpWWdY+PM7i5Bfv6jBnYT4NQ947x/buk7D9ztqSqZlC3VJ1O6+/t7ZRJ/+SkR+69u1dCfLWpTulnVYJQYao7c/6C3eFIS0my2GxpKck9E+ICvLxiQoOVENO0lKQdR08Yzea0lORuHeON9Q1HM7Nemjje39sgSVJ0SIivp0dOYbHd5nj1bxPc9VqrzT7zbxPuHdC3tKo6OTqyrNZYcKV0/B2DB3Tt4uvlGR4QsPXgoa4JCU+MGsH+h4E0/+tR4S9Nevw+o8JrNxJCW4+xf3FU+BtO9V/j3wZw/i4XuuFVbu6o8A8nFmsRaVx/KKFCitqkbIQIKRGhykQKaNGEFF2q9ZQFIVSZdEOtRB3lVK7GpS2nwggxwJR4ZdfsWytBi7IWKai5TMN1OpYSWY+u1bGaFaYWucQljLWesVF+VZ63TcK78gjwGnmMglaBCa1PpbjbzQurQHhDHavlSZuX9Gl9fuCqQPE/D6T5Synvt8oEbXvAX6qMEWcVJ9YfgsvlFa3zxq53OP7MtvaWu+e/CLF+pY0VH2XPidMLVq37JQ3wjzNm7N/d3n9ta29pA3zLEAteW2S2zatljIkIThs/VpnxZS1QoqMYY0azWXFjW293lVZr2Xw1obm5fstV4wFa73nNsAsASZZhK5Y3n7zlQq79qbKoU6s9XfdjNJuV9cldm2RCTA0NVw+81azXn51YypvYsGPnhyvWKMG413/FSujEz8dP1NY3MMZAy1qpzaETjDHG/j53QUlldevtoCUsArRE+kJ4dfjZOsoPAABhc+Cya5k4xhiR5Tc+++qht95576ulzy/8tDnur+VA1HIhV74yghChqzmrytUVTJv/cUFZOWw5CkJYUVM7fcHHpHmNu+YHuYWI9WdfVk4Rt0qqajKyc3p3SnTT6SxWK2PM6XRKMhFFQdmhymicuvDT2qamMf37Ekora42yLJfXGi1Wm6e7W31TY1JUVGRIUL25sclmNdY3VJlMPl6ex3PyvNzdMIRXysohQla7vcpodHfTIwhPZOdU1dUH+noDAKEyf2Cq9/Jwt9ntxaXlPl6eMiGiKB7IzKqzOT55cZrDaokKCxUwPnI2s76xycvdvai0XCbypZIyT71eFAUAQOb5CzV19QE+3oqZPZp5DiEky7JDklJioqNDghuamurMjTa7wylJCMLUuDg/bwOE8HRefn1jo5+3gRPr93RrIYRlNdXpp8+ezr9Y22Dec+x43uXLNVU1lbW1UR3CZEIwQtsOHh47qP93B4+MSOvlrtV+vHbDm998i2U6/bOvhvfuYTTV3ffW2+MGDzx9Pn/cm+9oGXt92SpTbe2mA+nrDx4edVvvMS+/4enu5pSk+1/7x9T7R7/2xTfl1dVXSsu8vbz8vLwopVa7/ePla/x8DBt27NLrdREhwUoA2fGc3MLKqqqqqqT42DB//2c++NhmsxdcLg7w9n7+06/2HT9RUWOcs3zVQ0OHHMnMSj95esn2nRCwpKjIyXPmqRGcs2y1u1rtqdePfX326AF9zxcWTXx3/v5jpyxWq9Vue2LhP/8+5t6dGcdPZOcu/mGzp5s+PrzDrbJk663hY9ntzo7RUbOefHTV7r33Dhp48FyuTOS0bqkAAAFjiZDNh44Cu7O+oWHzoQwGYaf4OL1W+9IjE0P9/c6cv5AcH6fT6hyy3CWxoyAKLz36ty4JsUCtWTnnjcyCIp1eHxYZLhGSGBstqFUQQIyFtfvS+/XoHhMaokTV+3h5zXr6sQfefDswKGhgj26EEiWqjjLmcDotNrvd4QQQMMTWHTw8qE/vuMiIkKCA+Jiod6c+0WCzbT58NCk6MjEqUqtS512+Qgk5mJXzxP1jGYKCRhMfGaHWah1OqXfnzhKhHz4/9fExIzvGRCMsMMbiIzokRoZjUbh4uaS1h8eJ9TuMuWRZLiop+3H3/s6REQkRHYz1DZXmJq1G45QkCOHp8/nJsZEhgf6Tht6xbv9BBCGj1NvdDQCgFUW1Wg0AwBhpVGpCiF6nAwAIEPkZvCilep0WAGBptIiiCBgDGMmy/PdR90wff9+oWXMulJRBCBlgDU1NHy5bvXD61ILLJceyshFERPHQAQgPCXrtsckJHcJKKivnPDz5oTsHD5v5hrG+AQLmrtUCAHRqUacSnl20uKCq2t/XByKEBWF03z6vLfps4qABo/rdxhgQRZVapZIp0el0kaEhOo2GEFkURYTQE/M/qm5s9PUy6LWaW2ic+GdfxV4x+90SE3y8PGrrGx66eygA4MHB/bsmdlT8X5vd/tnq9RPuHtq5U8d6m/WDFasPnjxdazSWVVXXGI21JlN5VVVVba2ppqboyhWZElN1TUVNTbXReKW8vKCktKG2tsZoHHFb70VrNxZeLmm0WDMys35MP3JH9y5dIjrIRFYWbLI5nPf0T+vWqWPX+Lii0jLFv6aUllVWZRYUr926/cfDR6bcN3rf0eODUzsnhQbLMtFotAfPntMiFOhlGNKt65rd+y9cKbE67KeuXLE0NjbarJnFJX5+vk02u7XWaKypKimvqGtoqK2uyr1wsUtifNGVUnNtjdFU56bX5RYVO+329NNnxg3q7+XldUtMMNx6Uzobdu65Ulb2/OSJAEIEodVuz714yc/HJyww4HJ5RYPZLIiiVq2y2R2Bvj4VtUYPvU6v05VWVPoYDAyw+npzgL9vbV29ShS0ao2pri4kMMDHYDhz/oLBw81ud3h5eiCMT2VlJ8fGhgYG/NqsOaV5hUUOSbbb7VjA3TolVhlN2fkXOsXGhAb4P7NwsYdGvKdP724d47VqtUxI+unMDgF+Dkmqqatfvn3n9AfuX7l915GcvE3vvXWlosLX2xtCWGs0+XobQgMDi8vKGxsbQ4OD3PS6g6cyY8OCm6y2qNAQjVp9a8hDtwSxXOKTgPHlikq9Vuvr5fk7frj/NvDBNRf+G/P4miyWMa/Njg4J+nzmC9cnRR4+l/Xy4q/efWRSRl5+VV39wheeBX85/BUmoV3yUrNvCyFs4URrZrT+o81uLo3KtZuiYvxb4lJ6tYAjagmaQAg1WazZFy9RxlLiY930+maaUqrEPCCECsvK8/IvePl4p3VOUb4Z1Or2XM/S+mYYYEq5JU6sP8p63erTHa0N5P8nR4gTqz3i+pVa2wi/Sr7/X3VpVk4sjnYskHJwYnFwcGJxcGJxcGJxcHBicXBicXBicXBicXBwYnFwYnFwYnFwcGJxcGJxcGJxcHBicXBicXBicXBwYnFwYnFwYnFwcGJxcGJxcGJxcHBicXBicXBicXBwYnFwYnFwYnFwcGJxcGJxcGJxcHBicXBicXBicXBwYnFwYnFwYnFwcGJxcGJxcGJxcHBicXBicXBicXBwYnH8GfB/fS2Du7R+qfAAAAAASUVORK5CYII=' style='height:60px;object-fit:contain;display:block;margin:0 auto 4px auto;' alt='HHA Group'><h2>COMPROBANTE DE PAGO</h2>"
        "<p><strong>HHA Group &mdash; Laboratorio Espagiria / &Aacute;NIMUS Lab</strong></p>"
        "<p>Per&iacute;odo: <strong>"+periodo+"</strong> &nbsp;|&nbsp; "
        "Estado: <span style='display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:700;"+badge+"'>"+estado+"</span>"
        +aprobado_txt+pagado_txt+"</p></div>"
        "<table>"
        "<tr><td class='lbl'>Nombre:</td><td><strong>"+nom+" "+ape+"</strong></td>"
        "<td class='lbl'>C&eacute;dula:</td><td><strong>"+ced+"</strong></td></tr>"
        "<tr><td class='lbl'>Cargo:</td><td>"+cargo+"</td><td class='lbl'>Empresa:</td><td>"+empresa+"</td></tr>"
        "<tr><td class='lbl'>D&iacute;as trabajados:</td><td>"+str(dias)+" d&iacute;as</td>"
        "<td class='lbl'>Riesgo ARL:</td><td>Nivel "+str(riesgo)+"</td></tr>"
        "</table>"
        "<div class='sec'>DEVENGADO</div><table>"
        "<tr><td class='lbl'>Salario base</td><td class='val'>"+cop(sal)+"</td></tr>"
        "<tr><td class='lbl'>Auxilio de transporte</td><td class='val'>"+cop(aux)+"</td></tr>"
        "<tr><td class='lbl'>Valor horas extras</td><td class='val'>"+cop(vhe)+"</td></tr>"
        "<tr><td class='lbl'>Bonificaciones</td><td class='val'>"+cop(bonos)+"</td></tr>"
        "<tr style='font-weight:700'><td>Total devengado</td><td class='val'>"+cop(sal+aux+vhe+bonos)+"</td></tr>"
        "</table>"
        "<div class='sec'>DEDUCCIONES</div><table>"
        "<tr><td class='lbl'>Salud (4%)</td><td class='val'>&minus;"+cop(ds)+"</td></tr>"
        "<tr><td class='lbl'>Pensi&oacute;n (4%)</td><td class='val'>&minus;"+cop(dp)+"</td></tr>"
        "<tr><td class='lbl'>Otros descuentos</td><td class='val'>&minus;"+cop(otros)+"</td></tr>"
        "<tr style='font-weight:700'><td>Total deducciones</td><td class='val'>&minus;"+cop(ds+dp+otros)+"</td></tr>"
        "</table>"
        "<div class='neto'>NETO A PAGAR: "+cop(neto)+"</div>"
        +(("<div class='sec' style='margin-top:12px;background:#15803d'>DATOS BANCARIOS PARA PAGO</div><table>"
        "<tr><td class='lbl'>Banco:</td><td><strong>"+(banco or "Sin registrar")+"</strong></td>"
        "<td class='lbl'>Tipo de cuenta:</td><td>"+(tipo_cta or "—")+"</td></tr>"
        "<tr><td class='lbl'>N&uacute;mero de cuenta:</td><td colspan='3' style='font-family:monospace;font-weight:700;color:#166534;font-size:13px;'>"+(num_cta or "—")+"</td></tr>"
        "</table>") if (banco or num_cta) else "")
        +"<div class='sec' style='margin-top:12px'>APORTES EMPLEADOR (informativo)</div><table>"
        "<tr><td class='lbl'>Salud (8.5%)</td><td class='val'>"+cop(ap_s)+"</td>"
        "<td class='lbl'>Pensi&oacute;n (12%)</td><td class='val'>"+cop(ap_p)+"</td></tr>"
        "<tr><td class='lbl'>ARL (Nivel "+str(riesgo)+")</td><td class='val'>"+cop(ap_arl)+"</td>"
        "<td class='lbl'>SENA (2%)</td><td class='val'>"+cop(ap_sena)+"</td></tr>"
        "<tr><td class='lbl'>ICBF (3%)</td><td class='val'>"+cop(ap_icbf)+"</td>"
        "<td class='lbl'>Caja comp. (4%)</td><td class='val'>"+cop(ap_caja)+"</td></tr>"
        "<tr style='font-weight:700'><td>Total aportes empleador</td><td class='val'>"+cop(ap_tot)+"</td>"
        "<td class='lbl'>Costo total empresa</td><td class='val' style='color:#6D28D9'>"+cop(neto+ap_tot)+"</td></tr>"
        "</table>"
        "<div class='footer'>Generado por Sistema HHA Group &nbsp;|&nbsp; "+periodo+"</div>"
        "<div style='text-align:center;margin-top:10px'>"
        "<button onclick='window.print()' style='padding:8px 20px;background:#1C1917;color:#fff;border:none;border-radius:6px;cursor:pointer'>&#128424; Imprimir</button>"
        "</div></body></html>"
    )
    return Response(html, mimetype="text/html")

@bp.route("/api/rrhh/nomina/<periodo>/aprobar", methods=["PATCH"])
def rrhh_nomina_aprobar(periodo):
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({"error": "Sin permiso"}), 403
    conn = get_db(); c = conn.cursor()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    c.execute("UPDATE nomina_registros SET estado='Aprobada',aprobado_por=?,aprobado_en=? WHERE periodo=?", (u, ts, periodo))
    updated = c.rowcount; conn.commit()
    return jsonify({"ok": True, "periodo": periodo, "aprobados": updated, "por": u})

@bp.route("/api/rrhh/nomina/<periodo>/pagar", methods=["PATCH"])
def rrhh_nomina_pagar(periodo):
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({"error": "Sin permiso"}), 403
    conn = get_db(); c = conn.cursor()
    # Only mark as Pagada if all records are already Aprobada
    c.execute("SELECT COUNT(*) FROM nomina_registros WHERE periodo=?", (periodo,))
    total = c.fetchone()[0]
    if total == 0:
        return jsonify({"error": "No hay registros para este período"}), 404
    c.execute("SELECT COUNT(*) FROM nomina_registros WHERE periodo=? AND estado != 'Aprobada'", (periodo,))
    no_aprobados = c.fetchone()[0]
    if no_aprobados > 0:
        return jsonify({"error": "La nómina debe estar aprobada antes de marcar como pagada"}), 400
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    c.execute("UPDATE nomina_registros SET estado='Pagada',pagado_por=?,pagado_en=? WHERE periodo=?", (u, ts, periodo))
    pagados = c.rowcount

    # ── Gap 4 cerrado: nomina → flujo_egresos automatico ──
    # Cada vez que se marca un periodo de nomina como Pagada, espejarlo en
    # flujo_egresos. Idempotente por (referencia=NOM-{periodo}). Si Sebastian
    # corre /pagar dos veces el mismo periodo, no se duplica el egreso.
    egresos_creados = 0
    try:
        # Total neto pagado del periodo (schema real: salario_neto)
        total_bruto_row = c.execute("""
            SELECT COALESCE(SUM(salario_neto),0),
                   COUNT(DISTINCT empleado_id)
            FROM nomina_registros WHERE periodo=?
        """, (periodo,)).fetchone()
        total_devengado = float(total_bruto_row[0] or 0)
        n_empleados = total_bruto_row[1] or 0

        if total_devengado > 0:
            ref = f'NOM-{periodo}'
            ya = c.execute("SELECT id FROM flujo_egresos WHERE referencia=?", (ref,)).fetchone()
            if not ya:
                fecha = datetime.now().strftime('%Y-%m-%d')
                # Periodo en flujo se interpreta como YYYY-MM. Si el periodo
                # de nomina viene como 2026-04 lo usamos directo; si viene
                # con dia tomamos los primeros 7 chars
                periodo_flujo = (periodo or fecha)[:7]
                concepto = f'Nomina {periodo} ({n_empleados} empleados)'
                c.execute("""
                    INSERT INTO flujo_egresos
                    (fecha, empresa, concepto, categoria, monto, periodo,
                     fuente, referencia, creado_por)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (fecha, 'ESPAGIRIA', concepto, 'Nomina',
                      total_devengado, periodo_flujo,
                      'nomina_auto', ref, u or 'sistema_sync'))
                egresos_creados = 1
    except Exception:
        # No bloquear el flujo de pago si el espejo falla
        pass

    conn.commit()
    return jsonify({"ok": True, "periodo": periodo, "pagados": pagados,
                    "por": u, "fecha": ts,
                    "egresos_flujo_creados": egresos_creados})

@bp.route("/api/rrhh/nomina/importar-excel", methods=["POST"])
def rrhh_nomina_importar():
    u = session.get("compras_user", "")
    if "file" not in request.files:
        return jsonify({"error": "No se recibio archivo"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400
    try:
        import openpyxl, io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(f.read()), data_only=True)
    except Exception as ex2:
        return jsonify({"error": "No se pudo leer: " + str(ex2)}), 400
    # Scan sheets newest-first for employee rows
    sheet_data = []
    for sh_name in reversed(wb.sheetnames):
        ws = wb[sh_name]
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = None
        for ri, row in enumerate(rows):
            if row[1] and str(row[1]).strip().upper() in ("C.C.", "C.C"):
                hdr_idx = ri; break
        if hdr_idx is None:
            continue
        for row in rows[hdr_idx + 2:]:
            num = row[0]; cc = row[1]; nombre = row[2]
            if not (isinstance(num, (int, float)) and cc and nombre):
                if row[0] is None and row[1] is None:
                    break
                continue
            dias = row[6] if len(row) > 6 and isinstance(row[6], (int, float)) else 15
            sheet_data.append({
                "cedula": str(int(cc)) if isinstance(cc, float) else str(cc).strip(),
                "dias": int(dias) if dias else 15,
            })
        if sheet_data:
            break
    if not sheet_data:
        return jsonify({"error": "No se encontraron datos en el archivo"}), 400
    conn = get_db(); c = conn.cursor()
    cedulas = [d["cedula"] for d in sheet_data]
    ph = ",".join("?" * len(cedulas))
    c.execute("SELECT id,cedula,nombre,apellido FROM empleados WHERE cedula IN ("+ph+")", cedulas)
    db_emps = {r[1]: {"id": r[0], "nombre": r[2]+" "+r[3]} for r in c.fetchall()}
    result = []
    for d in sheet_data:
        if d["cedula"] in db_emps:
            emp = db_emps[d["cedula"]]
            result.append({"empleado_id": emp["id"], "nombre": emp["nombre"],
                           "cedula": d["cedula"], "dias_trabajados": d["dias"]})
    return jsonify({"ok": True, "matched": len(result), "total_excel": len(sheet_data), "data": result})

@bp.route("/api/rrhh/admin/seed-bancos", methods=["POST"])
def rrhh_seed_bancos():
    """Admin: fuerza UPDATE de banco/cuenta en todos los empleados reales."""
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({"error": "Sin permiso"}), 403
    _bank_data = [
        ("16632635",   "BBVA",         "813000200051521",   "AHORROS"),
        ("1143874047", "DAVIVIENDA",   "10470059592",       "AHORROS"),
        ("1128724125", "BANCOLOMBIA",  "91220528389",       "AHORROS"),
        ("1026560691", "BANCOLOMBIA",  "91291991802",       "AHORROS"),
        ("1026560690", "BANCOLOMBIA",  "91246950747",       "AHORROS"),
        ("1109663762", "BANCOLOMBIA",  "6160474104",        "AHORROS"),
        ("1098307374", "DAVIVIENDA",   "0550488436467077",  "AHORROS DAMAS"),
        ("1097397765", "BANCOLOMBIA",  "91273689724",       "AHORROS"),
        ("1235252199", "BANCOLOMBIA",  "6107281001",        "AHORROS"),
        ("1006054219", "AV-VILLAS",    "148707529",         "AHORROS"),
        ("1007854652", "BANCOLOMBIA",  "91219764516",       "AHORROS"),
        ("1005875757", "BANCOLOMBIA",  "91219757421",       "AHORROS"),
        ("1007601298", "BANCOLOMBIA",  "3146792620",        "NEQUI"),
        ("43976397",   "BANCOLOMBIA",  "81583095349",       "AHORROS"),
        ("1143846075", "DAVIVIENDA",   "0570019170026397",  "AHORROS"),
        ("1044912921", "BANCOLOMBIA",  "80798012383",       "AHORROS"),
        ("1007932197", "DAVIVIENDA",   "0570488471748506",  "AHORROS"),
        ("14639995",   "BANCOLOMBIA",  "60566122726",       "AHORROS"),
        ("1193447691", "CAJA SOCIAL",  "24103175746",       "AHORROS"),
        ("1001937292", "BANCO BOGOTA", "164579443",         "AHORROS"),
    ]
    conn = get_db(); c = conn.cursor()
    actualizados = 0
    detalles = []
    for ced, banco, num, tipo in _bank_data:
        c.execute("UPDATE empleados SET banco=?, numero_cuenta=?, tipo_cuenta=? WHERE cedula=?",
                  (banco, num, tipo, ced))
        if c.rowcount > 0:
            actualizados += 1
            detalles.append(ced)
    conn.commit()
    return jsonify({"ok": True, "actualizados": actualizados, "cedulas": detalles})


# ════════════════════════════════════════════════════════════════════════
# RH AMPLIADO (29-abr-2026): documentos + eventos + llamados + compromisos
# Cumple ley colombiana 100/776 para incapacidades y accidentes laborales.
# ════════════════════════════════════════════════════════════════════════

def _calcular_pago_incapacidad(salario_mensual, tipo, dias):
    """Calcula el desglose legal de pago segun ley colombiana.

    Returns: dict con detalle por rango de dias, total empleador, total EPS,
    total ARL, y si descuenta o no de la nomina.

    Reglas (Ley 100, Ley 776, Ley 1822 maternidad):
    - Incapacidad COMUN (origen general, enfermedad no laboral):
        Dias 1-2: 100% empleador (no le paga EPS)
        Dias 3-90: 66.66% EPS (auxilio monetario)
        Dias 91-180: 50% EPS
        > 180: pension de invalidez (no aplicable a este calculo)
    - Incapacidad LABORAL / ACCIDENTE TRABAJO:
        Dias 1-2: 100% empleador
        Dias 3+: 100% ARL (toda la incapacidad)
    - Maternidad: 18 semanas (126 dias) al 100% por EPS
    - Paternidad: 14 dias al 100% por EPS
    - Luto: 5 dias habiles al 100% empleador (Ley 1280)
    """
    dias = max(0, int(dias or 0))
    salario_diario = float(salario_mensual or 0) / 30.0
    detalle = []
    pago_empleador = 0.0
    pago_eps = 0.0
    pago_arl = 0.0

    if tipo == 'incapacidad_comun':
        if dias >= 1:
            d1 = min(dias, 2); pago = round(d1 * salario_diario, 0)
            pago_empleador += pago
            detalle.append({'rango':'Día 1-2', 'pagador':'EMPLEADOR', 'pct':100, 'dias':d1, 'monto':pago})
        if dias > 2:
            d2 = min(dias-2, 88)  # dias 3-90 = 88 dias
            pago = round(d2 * salario_diario * 0.6666, 0)
            pago_eps += pago
            detalle.append({'rango':'Día 3-90', 'pagador':'EPS', 'pct':66.66, 'dias':d2, 'monto':pago})
        if dias > 90:
            d3 = min(dias-90, 90)  # dias 91-180 = 90 dias
            pago = round(d3 * salario_diario * 0.50, 0)
            pago_eps += pago
            detalle.append({'rango':'Día 91-180', 'pagador':'EPS', 'pct':50, 'dias':d3, 'monto':pago})
    elif tipo in ('incapacidad_laboral', 'accidente_trabajo'):
        if dias >= 1:
            d1 = min(dias, 2); pago = round(d1 * salario_diario, 0)
            pago_empleador += pago
            detalle.append({'rango':'Día 1-2', 'pagador':'EMPLEADOR', 'pct':100, 'dias':d1, 'monto':pago})
        if dias > 2:
            d2 = dias - 2
            pago = round(d2 * salario_diario, 0)
            pago_arl += pago
            detalle.append({'rango':'Día 3 en adelante', 'pagador':'ARL', 'pct':100, 'dias':d2, 'monto':pago})
    elif tipo == 'licencia_maternidad':
        # 18 semanas = 126 dias al 100% EPS
        d = min(dias, 126); pago = round(d * salario_diario, 0)
        pago_eps += pago
        detalle.append({'rango':f'Día 1-{d}', 'pagador':'EPS', 'pct':100, 'dias':d, 'monto':pago})
        if dias > 126:
            d_extra = dias - 126; pago_extra = round(d_extra * salario_diario, 0)
            pago_empleador += pago_extra
            detalle.append({'rango':f'Día {127}-{dias}', 'pagador':'EMPLEADOR (excede ley)', 'pct':100, 'dias':d_extra, 'monto':pago_extra})
    elif tipo == 'licencia_paternidad':
        d = min(dias, 14); pago = round(d * salario_diario, 0)
        pago_eps += pago
        detalle.append({'rango':f'Día 1-{d}', 'pagador':'EPS', 'pct':100, 'dias':d, 'monto':pago})
    elif tipo == 'licencia_luto':
        d = min(dias, 5); pago = round(d * salario_diario, 0)
        pago_empleador += pago
        detalle.append({'rango':f'Día 1-{d} (Ley 1280)', 'pagador':'EMPLEADOR', 'pct':100, 'dias':d, 'monto':pago})
    elif tipo == 'licencia_no_remunerada':
        detalle.append({'rango':f'{dias} días', 'pagador':'NO PAGA', 'pct':0, 'dias':dias, 'monto':0})
    elif tipo == 'permiso_remunerado':
        pago = round(dias * salario_diario, 0)
        pago_empleador += pago
        detalle.append({'rango':f'{dias} días', 'pagador':'EMPLEADOR', 'pct':100, 'dias':dias, 'monto':pago})
    # llamados/felicitacion/reinduccion no llevan pago

    total = pago_empleador + pago_eps + pago_arl
    # Descuento de nomina = lo que NO paga el empleador (recupera de EPS/ARL via reembolso)
    # Por simplicidad: si es incapacidad/accidente, los días 3+ se descuentan de
    # nomina y se reembolsan via EPS/ARL.
    descuento_nomina = pago_eps + pago_arl  # lo que se descuenta del pago directo

    return {
        'tipo': tipo, 'dias': dias,
        'salario_mensual': float(salario_mensual or 0),
        'salario_diario': round(salario_diario, 0),
        'detalle': detalle,
        'pago_empleador': pago_empleador,
        'pago_eps': pago_eps,
        'pago_arl': pago_arl,
        'descuento_nomina': descuento_nomina,
        'total': total,
    }


@bp.route('/api/rrhh/calcular-pago-evento', methods=['POST'])
def rrhh_calcular_pago_evento():
    """Preview del calculo legal de un evento (incapacidad/accidente/licencia).

    Body: {salario_mensual, tipo, dias}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    return jsonify(_calcular_pago_incapacidad(
        float(d.get('salario_mensual') or 0),
        (d.get('tipo') or '').strip(),
        int(d.get('dias') or 0),
    ))


@bp.route('/api/rrhh/eventos', methods=['GET', 'POST'])
def rrhh_eventos():
    """Lista eventos RH (todos o de un empleado) o crea uno nuevo.

    GET querystring: ?empleado_id=X&tipo=X&estado=X
    POST body: {empleado_id, tipo, fecha_inicio, fecha_fin, dias, ...}
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        emp_id = int(d.get('empleado_id') or 0)
        tipo = (d.get('tipo') or '').strip()
        if not emp_id or not tipo:
            return jsonify({'error': 'empleado_id y tipo requeridos'}), 400
        # Calcular dias si vienen fecha_inicio + fecha_fin
        f_ini = (d.get('fecha_inicio') or '').strip()
        f_fin = (d.get('fecha_fin') or '').strip()
        dias = int(d.get('dias') or 0)
        if not dias and f_ini and f_fin:
            try:
                from datetime import date
                d1 = date.fromisoformat(f_ini); d2 = date.fromisoformat(f_fin)
                dias = (d2 - d1).days + 1
            except Exception: pass

        # Calcular pago si aplica (necesita salario_base del empleado)
        emp = c.execute("SELECT salario_base FROM empleados WHERE id=?", (emp_id,)).fetchone()
        salario_mensual = float(emp[0] or 0) if emp else 0
        calc = _calcular_pago_incapacidad(salario_mensual, tipo, dias)
        import json as _json

        cur = c.execute("""
            INSERT INTO rh_eventos (
                empleado_id, tipo, fecha_inicio, fecha_fin, dias,
                descripcion, diagnostico, cie10, entidad_emisora, origen,
                motivo, severidad, jefe_id, jefe_nombre, area,
                salario_diario_referencia, pago_empleador, pago_eps, pago_arl,
                descuento_nomina, calculo_detalle_json,
                documento_url, estado, registrado_por
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'registrada', ?)
        """, (
            emp_id, tipo, f_ini, f_fin or None, dias,
            (d.get('descripcion') or '').strip(),
            (d.get('diagnostico') or '').strip(),
            (d.get('cie10') or '').strip(),
            (d.get('entidad_emisora') or '').strip(),
            (d.get('origen') or '').strip(),
            (d.get('motivo') or '').strip(),
            (d.get('severidad') or '').strip(),
            d.get('jefe_id'),
            (d.get('jefe_nombre') or user).strip(),
            (d.get('area') or '').strip(),
            calc['salario_diario'],
            calc['pago_empleador'],
            calc['pago_eps'],
            calc['pago_arl'],
            calc['descuento_nomina'],
            _json.dumps(calc['detalle']),
            (d.get('documento_url') or '').strip(),
            user
        ))
        conn.commit()
        return jsonify({'ok': True, 'evento_id': cur.lastrowid, 'calculo': calc})

    # GET
    where = []
    params = []
    emp_id = request.args.get('empleado_id')
    if emp_id:
        where.append("empleado_id=?"); params.append(int(emp_id))
    tipo = request.args.get('tipo')
    if tipo:
        where.append("tipo=?"); params.append(tipo)
    estado = request.args.get('estado')
    if estado:
        where.append("estado=?"); params.append(estado)
    sql = "SELECT * FROM rh_eventos"
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY fecha_inicio DESC, fecha_registro DESC LIMIT 500"
    rows = c.execute(sql, params).fetchall()
    cols = [x[0] for x in c.description]
    return jsonify({'eventos': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/rrhh/eventos/<int:evt_id>/aprobar', methods=['POST'])
def rrhh_evento_aprobar(evt_id):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    cur = c.execute("""
        UPDATE rh_eventos
           SET estado='aprobada', aprobado_por=?, fecha_aprobacion=datetime('now')
         WHERE id=? AND estado='registrada'
    """, (user, evt_id))
    if cur.rowcount == 0:
        return jsonify({'error': 'Evento no encontrado o ya aprobado'}), 404
    conn.commit()
    return jsonify({'ok': True, 'evento_id': evt_id})


@bp.route('/api/rrhh/eventos/<int:evt_id>/cerrar', methods=['POST'])
def rrhh_evento_cerrar(evt_id):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    obs = (d.get('observaciones') or '').strip()
    conn = get_db(); c = conn.cursor()
    c.execute("""
        UPDATE rh_eventos
           SET estado='cerrada', observaciones_cierre=?
         WHERE id=?
    """, (obs, evt_id))
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/rrhh/empleados/<int:emp_id>/timeline', methods=['GET'])
def rrhh_empleado_timeline(emp_id):
    """Timeline completo del empleado: eventos + capacitaciones + nomina."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    emp = c.execute("SELECT * FROM empleados WHERE id=?", (emp_id,)).fetchone()
    if not emp:
        return jsonify({'error': 'Empleado no encontrado'}), 404
    cols = [x[0] for x in c.description]
    empleado = dict(zip(cols, emp))

    eventos = c.execute(
        "SELECT * FROM rh_eventos WHERE empleado_id=? ORDER BY fecha_inicio DESC LIMIT 100",
        (emp_id,)
    ).fetchall()
    eventos_cols = [x[0] for x in c.description]
    eventos_list = [dict(zip(eventos_cols, r)) for r in eventos]

    # Resumen por tipo
    resumen = {}
    for ev in eventos_list:
        t = ev.get('tipo','otro')
        resumen[t] = resumen.get(t, 0) + 1

    documentos = c.execute(
        "SELECT id, tipo, nombre, fecha_emision, fecha_vencimiento FROM empleados_documentos WHERE empleado_id=? ORDER BY fecha_carga DESC",
        (emp_id,)
    ).fetchall()
    docs_cols = [x[0] for x in c.description]
    docs_list = [dict(zip(docs_cols, r)) for r in documentos]

    compromisos = c.execute(
        "SELECT * FROM rh_compromisos_mejora WHERE empleado_id=? ORDER BY fecha_creacion DESC LIMIT 50",
        (emp_id,)
    ).fetchall()
    comp_cols = [x[0] for x in c.description]
    comp_list = [dict(zip(comp_cols, r)) for r in compromisos]

    return jsonify({
        'empleado': empleado,
        'eventos': eventos_list,
        'documentos': docs_list,
        'compromisos': comp_list,
        'resumen_eventos': resumen,
    })


@bp.route('/api/rrhh/empleados/<int:emp_id>/documentos', methods=['GET', 'POST'])
def rrhh_documentos(emp_id):
    """GET: lista documentos del empleado. POST: cargar uno nuevo."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        cur = c.execute("""
            INSERT INTO empleados_documentos (
                empleado_id, tipo, nombre, archivo_url, archivo_data, mime_type,
                fecha_emision, fecha_vencimiento, observaciones, cargado_por
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            emp_id,
            (d.get('tipo') or 'otro').strip(),
            (d.get('nombre') or '').strip(),
            (d.get('archivo_url') or '').strip(),
            (d.get('archivo_data') or '').strip(),
            (d.get('mime_type') or '').strip(),
            (d.get('fecha_emision') or '').strip() or None,
            (d.get('fecha_vencimiento') or '').strip() or None,
            (d.get('observaciones') or '').strip(),
            user
        ))
        conn.commit()
        return jsonify({'ok': True, 'documento_id': cur.lastrowid})
    # GET
    rows = c.execute(
        "SELECT id, tipo, nombre, archivo_url, mime_type, fecha_emision, fecha_vencimiento, observaciones, cargado_por, fecha_carga FROM empleados_documentos WHERE empleado_id=? ORDER BY fecha_carga DESC",
        (emp_id,)
    ).fetchall()
    cols = [x[0] for x in c.description]
    return jsonify({'documentos': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/rrhh/llamados-atencion', methods=['GET', 'POST'])
def rrhh_llamados_atencion():
    """API conveniente para crear/listar llamados de atencion (subset de eventos).

    POST body: {empleado_id, severidad (verbal|escrito|suspension), motivo,
                area?, jefe_nombre?, descripcion?, fecha?, plan_mejora?}
    Crea evento + (opcional) compromiso_mejora.
    GET: lista llamados activos
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        emp_id = int(d.get('empleado_id') or 0)
        severidad = (d.get('severidad') or 'verbal').strip()
        tipo_map = {
            'verbal': 'llamado_atencion_verbal',
            'escrito': 'llamado_atencion_escrito',
            'suspension': 'suspension',
        }
        tipo = tipo_map.get(severidad, 'llamado_atencion_verbal')
        motivo = (d.get('motivo') or '').strip()
        if not emp_id or not motivo:
            return jsonify({'error': 'empleado_id y motivo requeridos'}), 400
        fecha = (d.get('fecha') or '').strip()
        if not fecha:
            from datetime import date
            fecha = date.today().isoformat()
        cur = c.execute("""
            INSERT INTO rh_eventos (
                empleado_id, tipo, fecha_inicio, dias, motivo, severidad,
                descripcion, jefe_nombre, area, registrado_por, estado
            ) VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?, 'registrada')
        """, (
            emp_id, tipo, fecha, motivo, severidad,
            (d.get('descripcion') or '').strip(),
            (d.get('jefe_nombre') or user).strip(),
            (d.get('area') or '').strip(),
            user
        ))
        evento_id = cur.lastrowid

        # Si plan_mejora viene, crear compromiso ligado
        compromiso_id = None
        plan = (d.get('plan_mejora') or '').strip()
        if plan:
            cur2 = c.execute("""
                INSERT INTO rh_compromisos_mejora (
                    empleado_id, evento_origen_id, titulo, descripcion,
                    tipo, plan_accion, fecha_compromiso, fecha_objetivo,
                    jefe_responsable, creado_por
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                emp_id, evento_id,
                f"Plan de mejora — {motivo[:80]}",
                f"Tras {tipo}: {motivo}",
                'reinduccion',
                plan,
                fecha,
                (d.get('fecha_objetivo') or '').strip() or None,
                (d.get('jefe_nombre') or user).strip(),
                user
            ))
            compromiso_id = cur2.lastrowid
        conn.commit()
        return jsonify({'ok': True, 'evento_id': evento_id, 'compromiso_id': compromiso_id})
    # GET
    rows = c.execute("""
        SELECT id, empleado_id, tipo, fecha_inicio, motivo, severidad,
               descripcion, jefe_nombre, area, estado, registrado_por
        FROM rh_eventos
        WHERE tipo IN ('llamado_atencion_verbal','llamado_atencion_escrito','suspension')
        ORDER BY fecha_inicio DESC LIMIT 200
    """).fetchall()
    cols = [x[0] for x in c.description]
    return jsonify({'llamados': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/rrhh/compromisos-mejora', methods=['GET', 'POST'])
def rrhh_compromisos_mejora():
    """GET: lista compromisos. POST: crear manual."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        cur = c.execute("""
            INSERT INTO rh_compromisos_mejora (
                empleado_id, evento_origen_id, titulo, descripcion, tipo,
                plan_accion, fecha_compromiso, fecha_objetivo,
                video_url, jefe_responsable, creado_por
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            int(d.get('empleado_id') or 0),
            d.get('evento_origen_id'),
            (d.get('titulo') or '').strip(),
            (d.get('descripcion') or '').strip(),
            (d.get('tipo') or 'reinduccion').strip(),
            (d.get('plan_accion') or '').strip(),
            (d.get('fecha_compromiso') or '').strip() or None,
            (d.get('fecha_objetivo') or '').strip() or None,
            (d.get('video_url') or '').strip(),
            (d.get('jefe_responsable') or user).strip(),
            user
        ))
        conn.commit()
        return jsonify({'ok': True, 'compromiso_id': cur.lastrowid})
    # GET
    estado = request.args.get('estado')
    where = ""
    params = ()
    if estado:
        where = "WHERE estado=?"
        params = (estado,)
    rows = c.execute(f"""
        SELECT * FROM rh_compromisos_mejora {where}
        ORDER BY fecha_creacion DESC LIMIT 200
    """, params).fetchall()
    cols = [x[0] for x in c.description]
    return jsonify({'compromisos': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/rrhh/compromisos-mejora/<int:cid>/completar', methods=['POST'])
def rrhh_compromiso_completar(cid):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    c.execute("""
        UPDATE rh_compromisos_mejora SET
          estado='completado',
          verificado_por=?,
          fecha_verificacion=datetime('now'),
          evidencia_url=?
        WHERE id=?
    """, (user, (d.get('evidencia_url') or '').strip(), cid))
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/rrhh/dashboard-rh-completo', methods=['GET'])
def rrhh_dashboard_completo():
    """KPIs ampliados de RH: empleados activos, eventos del mes, llamados,
    compromisos vencidos, vencimientos de documentos.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    def _safe(q, default=0):
        try:
            r = c.execute(q).fetchone()
            return r[0] if r and r[0] is not None else default
        except Exception:
            return default

    emp_activos = _safe("SELECT COUNT(*) FROM empleados WHERE COALESCE(estado,'Activo')='Activo'")
    eventos_mes = _safe("""SELECT COUNT(*) FROM rh_eventos
                           WHERE substr(fecha_inicio,1,7) = strftime('%Y-%m','now')""")
    incapac_activas = _safe("""SELECT COUNT(*) FROM rh_eventos
                               WHERE tipo IN ('incapacidad_comun','incapacidad_laboral')
                                 AND estado IN ('registrada','aprobada')
                                 AND date('now') BETWEEN fecha_inicio AND COALESCE(fecha_fin,'9999-12-31')""")
    llamados_30d = _safe("""SELECT COUNT(*) FROM rh_eventos
                            WHERE tipo IN ('llamado_atencion_verbal','llamado_atencion_escrito','suspension')
                              AND fecha_inicio >= date('now','-30 days')""")
    compromisos_pendientes = _safe("""SELECT COUNT(*) FROM rh_compromisos_mejora
                                      WHERE estado IN ('pendiente','en_progreso')""")
    docs_por_vencer = _safe("""SELECT COUNT(*) FROM empleados_documentos
                               WHERE COALESCE(fecha_vencimiento,'') != ''
                                 AND fecha_vencimiento >= date('now')
                                 AND fecha_vencimiento <= date('now','+30 days')""")
    docs_vencidos = _safe("""SELECT COUNT(*) FROM empleados_documentos
                             WHERE COALESCE(fecha_vencimiento,'') != ''
                               AND fecha_vencimiento < date('now')""")
    return jsonify({
        'empleados_activos': emp_activos,
        'eventos_mes': eventos_mes,
        'incapacidades_activas': incapac_activas,
        'llamados_atencion_30d': llamados_30d,
        'compromisos_pendientes': compromisos_pendientes,
        'documentos_por_vencer': docs_por_vencer,
        'documentos_vencidos': docs_vencidos,
    })


# ═══════════════════════════════════════════════════════
#  CALIDAD BPM — Página + API
# ═══════════════════════════════════════════════════════
