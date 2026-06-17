# blueprints/inventario.py — extraído de index.py (Fase C)
import os
import json
import sqlite3
import hmac
import time
from datetime import datetime, timedelta
from database import db_connect
from flask import Blueprint, jsonify, request, Response, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from config import DB_PATH, COMPRAS_USERS, ADMIN_USERS, CONTADORA_USERS, CALIDAD_USERS
from database import get_db
from auth import _client_ip, _is_locked, _record_failure, _clear_attempts, _log_sec
from audit_helpers import audit_log
from http_helpers import validate_money
from templates_py.rrhh_html import RRHH_HTML
from templates_py.compromisos_html import COMPROMISOS_HTML
from templates_py.home_html import HOME_HTML
from templates_py.hub_html import HUB_HTML
from templates_py.clientes_html import CLIENTES_HTML
from templates_py.calidad_html import CALIDAD_HTML
from templates_py.gerencia_html import GERENCIA_HTML
from templates_py.financiero_html import FINANCIERO_HTML
from templates_py.login_html import LOGIN_HTML
from templates_py.compras_html import COMPRAS_HTML
from templates_py.recepcion_html import RECEPCION_HTML
from templates_py.salida_html import SALIDA_HTML
from templates_py.solicitudes_html import SOLICITUDES_HTML
from templates_py.dashboard_html import DASHBOARD_HTML

bp = Blueprint('inventario', __name__)


# ── Helpers de permisos granulares ────────────────────────────────────────────
# Cualquier user autenticado puede LEER inventario (necesario para que el
# equipo vea stock/lotes desde su módulo). Pero las ESCRITURAS y operaciones
# críticas se restringen al rol correspondiente.
#
# Patrón de uso al inicio del endpoint:
#     u, err, code = _require_planta_write()
#     if err: return err, code

QC_USERS = CALIDAD_USERS | ADMIN_USERS


def _require_session():
    """Solo autenticación (cualquier user logueado). Usado en lecturas."""
    u = session.get('compras_user', '')
    if not u:
        return None, jsonify({'error': 'No autenticado'}), 401
    return u, None, None


def _require_planta_write():
    """Operaciones que ESCRIBEN inventario: cualquier usuario autenticado.

    Política decidida por CEO 2026-04-27: el flujo operativo cruza áreas
    (contadoría que registra recepción, asistentes que ajustan conteos,
    técnica que da entrada a piezas), por lo que se abre a cualquier
    usuario logueado. La trazabilidad se mantiene por:
      - audit_log con username + IP en cada operación
      - campo 'operador' en cada movimiento de la tabla movimientos
      - security_events con login/logout

    Operaciones destructivas (reset, delete masivo) siguen requiriendo
    _require_admin(). QC (aprobar lote, etc.) sigue requiriendo
    _require_qc() (CALIDAD + ADMIN).
    """
    u = session.get('compras_user', '')
    if not u:
        return None, jsonify({'error': 'No autenticado'}), 401
    return u, None, None


def _require_qc():
    """Para operaciones de QC: CALIDAD + ADMIN únicamente."""
    u = session.get('compras_user', '')
    if not u:
        return None, jsonify({'error': 'No autenticado'}), 401
    if u not in QC_USERS:
        return None, jsonify({
            'error': 'Solo equipo de Calidad o admins',
            'detail': f"User '{u}' no está en CALIDAD/ADMIN"
        }), 403
    return u, None, None


def _require_admin():
    """Para operaciones destructivas: solo ADMIN."""
    u = session.get('compras_user', '')
    if not u:
        return None, jsonify({'error': 'No autenticado'}), 401
    if u not in ADMIN_USERS:
        return None, jsonify({'error': 'Solo administradores'}), 403
    return u, None, None


def _validar_e_sign(c, signature_id, *, record_table, record_id, meaning, signer_username):
    """Valida una firma electrónica Part 11 (§11.200) sobre un registro.

    La firma debe existir en e_signatures, ser del usuario actual, sobre este
    registro exacto y con el meaning correcto. Mismo patrón que brd.py
    (_validar_signature). Devuelve True si la firma es válida.
    """
    if not signature_id:
        return False
    try:
        sig = c.execute(
            """SELECT id FROM e_signatures
               WHERE id=? AND record_table=? AND record_id=?
                 AND meaning=? AND signer_username=?""",
            (int(signature_id), record_table, str(record_id), meaning, signer_username),
        ).fetchone()
        return sig is not None
    except Exception:
        return False


def _mee_stock_real(c, codigo_mee):
    """MEE-FIX · 22-may-2026 · stock CANONICAL desde SUM(movimientos_mee).

    Reemplaza lectura de cache `maestro_mee.stock_actual` que podía
    drifear bajo carga. Ahora: una sola fuente de verdad = SUM movimientos.

    Args:
        c: cursor SQLite/PG
        codigo_mee: código del envase

    Returns:
        float · stock real (siempre >= 0)
    """
    try:
        r = c.execute(
            """SELECT COALESCE(SUM(CASE
                   WHEN tipo='Entrada' THEN cantidad
                   WHEN tipo='Salida'  THEN -cantidad
                   WHEN tipo='Ajuste'  THEN cantidad
                   ELSE 0 END), 0)
               FROM movimientos_mee
               WHERE mee_codigo=? AND COALESCE(anulado,0)=0""",
            (codigo_mee,),
        ).fetchone()
        return max(float(r[0] or 0), 0)
    except Exception:
        return 0.0

# ── Sebastian 5-may-2026: revisar stock_minimo desde Planta ────────
# El equipo (no solo admin) necesita ver si los stock_minimo configurados
# son reales antes de fiarse de las alertas. Endpoint read-only que
# reutiliza la logica de /api/admin/auditar-minimos.
@bp.route('/api/planta/auditar-minimos', methods=['GET'])
def planta_auditar_minimos():
    """Audit read-only de stock_minimo · accesible para todo Planta.

    Reutiliza _compute_audit_minimos() del modulo admin · misma logica
    pero sin requerir admin role (cualquier user autenticado puede ver).

    El APPLY sigue siendo solo admin via /api/admin/aplicar-minimos.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    try:
        horizonte = max(30, min(int(request.args.get('proyeccion_dias', 90)), 180))
    except (ValueError, TypeError):
        horizonte = 90
    # Sebastián 20-may-2026: dias_cobertura_minimo activa modo uniforme.
    dias_cob_raw = request.args.get('dias_cobertura_minimo')
    dias_cob = None
    if dias_cob_raw:
        try:
            dias_cob = max(7, min(int(dias_cob_raw), 365))
        except (ValueError, TypeError):
            dias_cob = None
    try:
        from blueprints.admin import _compute_audit_minimos
    except ImportError:
        from api.blueprints.admin import _compute_audit_minimos
    try:
        data = _compute_audit_minimos(horizonte, dias_cob)
        return jsonify(data)
    except Exception as e:
        __import__('logging').getLogger('inventario').warning(
            "planta_auditar_minimos falló: %s", e,
        )
        return jsonify({
            'error': 'No se pudo calcular el audit',
            'detalle': str(e)[:200],
        }), 500


@bp.route('/api/inventario')
def get_inventario():
    """Endpoint principal del Dashboard de Planta. Devuelve KPIs agrupados
    en 3 zonas: AHORA (acción hoy), CERCA (próximos 7-30d), CONTEXTO (totales).

    Sebastian (28-abr-2026): replanteo del dashboard. Antes solo devolvia
    4 KPIs sueltos. Ahora estructurado para poder ver de un vistazo qué
    requiere acción YA, qué viene pronto, y dónde estamos parados.
    """
    conn = get_db()
    c = conn.cursor()

    def _safe(query, default=0):
        # Sebastian 5-may-2026 (audit zero-error dashboard): ANTES capturaba
        # silenciosamente cualquier excepción y devolvía default=0 — ocultaba
        # bugs de DB locked, syntax errors, columns missing, etc. Ahora se
        # loguea con contexto · sigue retornando default para no romper UI
        # pero deja rastro auditable.
        try:
            r = c.execute(query).fetchone()
            return r[0] if r and r[0] is not None else default
        except Exception as _e:
            __import__('logging').getLogger('inventario').warning(
                "_safe query falló · default=%s · err=%s · query=%s",
                default, _e, query[:120].replace('\n', ' '),
            )
            return default

    # ── CONTEXTO (totales / composición) ──────────────────────────────
    mov = _safe('SELECT COUNT(*) FROM movimientos')
    prod_historico = _safe('SELECT COUNT(*) FROM producciones')
    stock_total = _safe("SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END),0) FROM movimientos")
    alrt = _safe('SELECT COUNT(*) FROM alertas')

    # ── AHORA (crítico — acción hoy) ──────────────────────────────────
    # MPs sin stock (en cero)
    # FIX 30-may-2026 · audit Planta/Dashboard · estos KPIs ("bloquean
    # producción") deben usar stock DISPONIBLE, no físico total. Antes sumaban
    # todos los movimientos incluyendo lotes en CUARENTENA/VENCIDO/RECHAZADO/
    # AGOTADO → un MP cuyo stock está atrapado en cuarentena salía "OK" aunque
    # no se puede producir con él. Ahora se excluyen esos estados (patrón
    # canónico de stock disponible · zero-error protocol). Las Salidas
    # (estado_lote NULL) siguen restando.
    _avail_filter = ("(estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) "
                     "NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))")
    mps_sin_stock = _safe(f"""
        SELECT COUNT(*) FROM maestro_mps m
        LEFT JOIN (SELECT material_id, SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock
                   FROM movimientos WHERE {_avail_filter} GROUP BY material_id) s ON m.codigo_mp=s.material_id
        WHERE m.activo=1 AND m.stock_minimo>0 AND COALESCE(s.stock,0) <= 0
    """)
    # MPs bajo mínimo (incluye sin stock) · mismo criterio de disponible
    mps_bajo_min = _safe(f"""
        SELECT COUNT(*) FROM maestro_mps m
        LEFT JOIN (SELECT material_id, SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock
                   FROM movimientos WHERE {_avail_filter} GROUP BY material_id) s ON m.codigo_mp=s.material_id
        WHERE m.activo=1 AND m.stock_minimo>0 AND COALESCE(s.stock,0)<m.stock_minimo
    """)
    # Lotes vencidos · Sebastian 5-may-2026 (audit zero-error dashboard):
    # ANTES usaba estado_lote='VENCIDO' estatico que NO se actualiza
    # automaticamente cuando un lote vence (queda como 'VIGENTE' aunque
    # ya pasó fecha_vencimiento). Drift critico: KPI mostraba 0 vencidos
    # mientras Bodega MP tenia 50 lotes vencidos visibles.
    # Fix: calcular dinamico desde fecha_vencimiento agrupando por lote
    # con stock > 0 (mismo pattern que /api/dashboard-stats compras.py:284).
    lotes_vencidos = _safe("""
        SELECT COUNT(*) FROM (
          SELECT material_id, lote,
                 MIN(fecha_vencimiento) as venc,
                 SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_g
          FROM movimientos
          WHERE COALESCE(lote,'') != ''
            AND fecha_vencimiento IS NOT NULL
            AND fecha_vencimiento != ''
          GROUP BY material_id, lote
          HAVING stock_g > 0 AND venc < date('now', '-5 hours')
        )
    """)

    # ── CERCA (próximos 7-30 días) ────────────────────────────────────
    # Próximas a producir: programadas activas en próximos 60 días
    prod_proximas = _safe("""
        SELECT COUNT(*) FROM produccion_programada
        WHERE LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
          AND fecha_programada >= date('now', '-5 hours', '-1 day')
          AND fecha_programada <= date('now', '-5 hours', '+60 day')
    """)
    # Lotes en cuarentena (esperando QC) · case-insensitive: calidad.py
    # escribe 'Cuarentena' (Capitalized), inventario.py escribe 'CUARENTENA'
    # (uppercase). UPPER normaliza ambos.
    # FIX 30-may-2026 · contar LOTES distintos, no filas de movimiento (un lote
    # con varias Entradas en cuarentena se contaba múltiples veces).
    lotes_cuarentena = _safe("""
        SELECT COUNT(*) FROM (
          SELECT DISTINCT material_id, lote FROM movimientos
          WHERE tipo='Entrada' AND COALESCE(lote,'') != ''
            AND UPPER(COALESCE(estado_lote,'')) IN ('CUARENTENA','CUARENTENA_EXTENDIDA')
        )
    """)
    # Vencimientos críticos próximos 30d · Sebastian 5-may-2026 (audit
    # zero-error): ANTES usaba estado_lote IN ('CRITICO','PROXIMO') que
    # NUNCA se asignan en DB (solo VIGENTE/CUARENTENA/VENCIDO se escriben).
    # KPI siempre retornaba 0 aunque hubiera lotes a 15 dias de vencer.
    # Fix: calcular dinamico desde fecha_vencimiento.
    venc_criticos = _safe("""
        SELECT COUNT(*) FROM (
          SELECT material_id, lote,
                 MIN(fecha_vencimiento) as venc,
                 SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_g
          FROM movimientos
          WHERE COALESCE(lote,'') != ''
            AND fecha_vencimiento IS NOT NULL
            AND fecha_vencimiento != ''
          GROUP BY material_id, lote
          HAVING stock_g > 0
             AND venc >= date('now', '-5 hours')
             AND venc <= date('now', '-5 hours', '+30 day')
        )
    """)
    # OCs en tránsito · Autorizada/Pagada sin recepción, MÁS Parcial (recepción
    # parcial · el resto sigue por llegar · FIX 30-may-2026 · antes no contaba
    # las Parcial porque ya tienen fecha_recepcion de la recepción parcial).
    ocs_transito = _safe("""
        SELECT COUNT(*) FROM ordenes_compra
        WHERE COALESCE(categoria,'') NOT IN
              ('Influencer/Marketing Digital','Cuenta de Cobro','SVC')
          AND (
                estado='Parcial'
             OR (estado IN ('Autorizada','Pagada')
                 AND (fecha_recepcion IS NULL OR fecha_recepcion = ''))
          )
    """)
    # MEE bajo mínimo (envases)
    # FIX · 21-may-2026 · maestro_mee usa `estado` no `activo`
    # FIX 30-may-2026 · usar stock canónico SUM(movimientos_mee) en vez del
    # cache stock_actual (que deriva · drift histórico MEE). anulado=0.
    mees_bajo_min = _safe("""
        SELECT COUNT(*) FROM maestro_mee m
        WHERE COALESCE(m.estado,'Activo')='Activo' AND COALESCE(m.stock_minimo,0)>0
          AND COALESCE((
                SELECT SUM(CASE WHEN tipo IN ('Entrada','Ajuste +','Ajuste') THEN cantidad
                                WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad ELSE 0 END)
                FROM movimientos_mee mm
                WHERE mm.mee_codigo=m.codigo AND COALESCE(mm.anulado,0)=0
              ),0) < COALESCE(m.stock_minimo,0)
    """)

    return jsonify({
        # Compatibilidad con frontend viejo
        'total_items': mov, 'movimientos': mov,
        'producciones': prod_historico,
        'producciones_historico': prod_historico,
        'producciones_proximas': prod_proximas,
        'alertas': alrt,
        'stock_total': round(stock_total, 2),
        # Sebastián 16-jun · interruptor recepción auto-VIGENTE (para que la UI
        # destilde la casilla "cuarentena" cuando está encendido).
        'recepcion_auto_vigente': (lambda: __import__('database').recepcion_auto_vigente())(),
        # KPIs nuevos del dashboard replanteado (zonas)
        'kpis': {
            'ahora': {
                'mps_sin_stock':   mps_sin_stock,
                'mps_bajo_minimo': mps_bajo_min,
                'lotes_vencidos':  lotes_vencidos,
            },
            'cerca': {
                'prod_proximas':       prod_proximas,
                'lotes_cuarentena':    lotes_cuarentena,
                'venc_criticos_30d':   venc_criticos,
                'ocs_en_transito':     ocs_transito,
                'mees_bajo_minimo':    mees_bajo_min,
            },
            'contexto': {
                'stock_total_g':   round(stock_total, 0),
                'lotes_bodega':    mov,
                'prod_historico':  prod_historico,
            },
        },
    })

@bp.route('/api/dashboard/insights', methods=['GET'])
def dashboard_insights():
    """Sprint Dashboard PRO #2 · Sebastián 20-may-2026.

    Consolida 3 widgets nuevos del Dashboard en una sola llamada:

    1. **planta_ahora** · producciones en curso, salas (libres/ocupadas/sucias),
       próxima programada · operario más activo.
    2. **mes_actual** · producciones del mes (terminadas / programadas) +
       kg producidos · sirve para mostrar progreso de mes vs estimado.
    3. **stats_extra** · MPs total activas, formula activas, operarios activos
       · contexto para Sebastián/Alejandro.

    Reusa queries existentes · no muta nada · cacheable a futuro.
    """
    from datetime import date as _date, datetime as _dt, timedelta as _td
    conn = get_db()
    c = conn.cursor()

    def _scalar(query, params=(), default=0):
        try:
            r = c.execute(query, params).fetchone()
            return r[0] if r and r[0] is not None else default
        except Exception as _e:
            __import__('logging').getLogger('inventario').warning(
                'dashboard_insights scalar fallo: %s · q=%s', _e,
                query[:80].replace('\n', ' '))
            return default

    # 1) Planta AHORA
    hoy = _date.today().isoformat()
    prod_en_curso = _scalar("""
        SELECT COUNT(*) FROM produccion_programada
        WHERE inicio_real_at IS NOT NULL
          AND fin_real_at IS NULL
          AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
    """)
    salas_libres = _scalar(
        "SELECT COUNT(*) FROM areas_planta WHERE COALESCE(activo,1)=1 "
        "AND COALESCE(estado,'libre')='libre'")
    salas_ocupadas = _scalar(
        "SELECT COUNT(*) FROM areas_planta WHERE COALESCE(activo,1)=1 "
        "AND COALESCE(estado,'')='ocupada'")
    salas_sucias = _scalar(
        "SELECT COUNT(*) FROM areas_planta WHERE COALESCE(activo,1)=1 "
        "AND COALESCE(estado,'')='sucia'")
    salas_total = _scalar(
        "SELECT COUNT(*) FROM areas_planta WHERE COALESCE(activo,1)=1")
    # Próxima producción programada (>= hoy)
    prox_row = c.execute(
        """SELECT producto, fecha_programada, COALESCE(cantidad_kg,0)
           FROM produccion_programada
           WHERE fecha_programada >= date('now', '-5 hours')
             AND inicio_real_at IS NULL
             AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
           ORDER BY fecha_programada ASC, id ASC LIMIT 1""",
    ).fetchone()
    proxima = None
    if prox_row:
        proxima = {'producto': prox_row[0] or '',
                   'fecha': prox_row[1] or '',
                   'kg': float(prox_row[2] or 0)}
    # Operarios con tarea hoy (sirve para "X operarios trabajando")
    operarios_activos = _scalar("""
        SELECT COUNT(DISTINCT op_id) FROM (
            SELECT operario_dispensacion_id AS op_id FROM produccion_programada
              WHERE date(fecha_programada)=? AND operario_dispensacion_id IS NOT NULL
            UNION
            SELECT operario_elaboracion_id FROM produccion_programada
              WHERE date(fecha_programada)=? AND operario_elaboracion_id IS NOT NULL
            UNION
            SELECT operario_envasado_id FROM produccion_programada
              WHERE date(fecha_programada)=? AND operario_envasado_id IS NOT NULL
            UNION
            SELECT operario_acondicionamiento_id FROM produccion_programada
              WHERE date(fecha_programada)=? AND operario_acondicionamiento_id IS NOT NULL
        )
    """, (hoy, hoy, hoy, hoy))

    # 2) Mes actual
    hoy_dt = _dt.now()
    primer_dia_mes = hoy_dt.replace(day=1).date().isoformat()
    fin_mes = (hoy_dt.replace(day=28) + _td(days=4)).replace(day=1).date().isoformat()
    prod_mes_completadas = _scalar("""
        SELECT COUNT(*) FROM produccion_programada
        WHERE LOWER(COALESCE(estado,''))='completado'
          AND fecha_programada >= ? AND fecha_programada < ?
    """, (primer_dia_mes, fin_mes))
    prod_mes_programadas = _scalar("""
        SELECT COUNT(*) FROM produccion_programada
        WHERE LOWER(COALESCE(estado,'')) NOT IN ('cancelado')
          AND fecha_programada >= ? AND fecha_programada < ?
    """, (primer_dia_mes, fin_mes))
    kg_producidos_mes = _scalar("""
        SELECT COALESCE(SUM(COALESCE(kg_real, cantidad_kg, 0)), 0)
        FROM produccion_programada
        WHERE LOWER(COALESCE(estado,''))='completado'
          AND fecha_programada >= ? AND fecha_programada < ?
    """, (primer_dia_mes, fin_mes), default=0.0)

    # 3) Stats extra
    mps_activas = _scalar(
        "SELECT COUNT(*) FROM maestro_mps WHERE COALESCE(activo,1)=1")
    formulas_activas = _scalar(
        "SELECT COUNT(*) FROM formula_headers WHERE COALESCE(activo,1)=1")
    operarios_pool = _scalar(
        "SELECT COUNT(*) FROM operarios_planta WHERE COALESCE(activo,1)=1")

    return jsonify({
        'planta_ahora': {
            'produciendo_ahora': prod_en_curso,
            'salas_libres': salas_libres,
            'salas_ocupadas': salas_ocupadas,
            'salas_sucias': salas_sucias,
            'salas_total': salas_total,
            'proxima_produccion': proxima,
            'operarios_con_tarea_hoy': operarios_activos,
        },
        'mes_actual': {
            'producciones_completadas': prod_mes_completadas,
            'producciones_programadas': prod_mes_programadas,
            'progreso_pct': round(
                (prod_mes_completadas / prod_mes_programadas * 100)
                if prod_mes_programadas > 0 else 0, 1),
            'kg_producidos': round(float(kg_producidos_mes or 0), 1),
            'mes': hoy_dt.strftime('%Y-%m'),
        },
        'stats_extra': {
            'mps_activas': mps_activas,
            'formulas_activas': formulas_activas,
            'operarios_pool': operarios_pool,
        },
    })


@bp.route('/api/formulas', methods=['GET', 'POST'])
def handle_formulas():
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        # Sebastián 12-may-2026: bugs criticos resueltos:
        #   1. INSERT items olvidaba cantidad_g_por_lote → fórmulas quedaban
        #      GPL_NO_SEMBRADO. Ahora calcula automaticamente.
        #   2. Trigger FK (mig 98) podia abortar items con material_id sin
        #      registrar en maestro_mps activo. Antes 500 sin rollback ·
        #      header zombie + items vacios. Ahora rollback + 400 + detalle.
        #   3. Tambien sembramos lote_size_kg = unidad_base_g/1000.
        data = request.json or {}
        prod = (data.get('producto_nombre') or '').strip()
        if not prod:
            return jsonify({'error': 'producto_nombre vacío'}), 400
        try:
            unidad_base_g = float(data.get('unidad_base_g', 1000) or 1000)
        except (ValueError, TypeError):
            return jsonify({'error': 'unidad_base_g debe ser numérico'}), 400
        if unidad_base_g <= 0:
            return jsonify({'error': 'unidad_base_g debe ser > 0'}), 400
        lote_size_kg = round(unidad_base_g / 1000.0, 3)
        items_input = data.get('items', [])
        if not items_input:
            return jsonify({'error': 'items vacío'}), 400

        # Validar items antes del INSERT (mejor rechazar early que rollback)
        items_validados = []
        for it in items_input:
            mid = (it.get('material_id') or '').strip()
            nm = (it.get('material_nombre') or '').strip()
            try:
                pct = float(it.get('porcentaje', 0) or 0)
            except (ValueError, TypeError):
                pct = 0
            if not mid or not nm or pct <= 0:
                continue
            # 27-may-2026 audit · un ingrediente no puede ser >100% de la fórmula
            # · cap lógico evita cantidad_g_por_lote absurda (riesgo INVIMA dosis).
            if pct > 100:
                return jsonify({
                    'error': f'Porcentaje inválido: {pct}% para {nm}',
                    'detalle': ('Un ingrediente no puede superar 100% de la '
                                'fórmula. Revisá el valor capturado.'),
                    'material_id_invalido': mid,
                }), 400
            # Validar que material_id existe en maestro_mps activo (FK preview)
            row = c.execute(
                "SELECT 1 FROM maestro_mps WHERE codigo_mp=? AND activo=1",
                (mid,)
            ).fetchone()
            if not row:
                return jsonify({
                    'error': f'MP no existe o está inactiva: {mid}',
                    'detalle': (f'La MP "{mid}" ({nm}) no está registrada en '
                                'maestro_mps con activo=1. Créala primero en '
                                'Bodega MP antes de usarla en una fórmula.'),
                    'material_id_invalido': mid,
                }), 400
            # cantidad_g_por_lote = (pct/100) × unidad_base_g
            gpl = round((pct / 100.0) * unidad_base_g, 2)
            items_validados.append({
                'material_id': mid,
                'material_nombre': nm,
                'porcentaje': pct,
                'cantidad_g_por_lote': gpl,
            })
        if not items_validados:
            return jsonify({'error': 'items todos inválidos (id, nombre y porcentaje>0 requeridos)'}), 400

        # FIX 1-jun-2026 audit fórmulas (P0-1) · CONSOLIDAR material_id duplicado ·
        # formula_items no tiene UNIQUE → dos filas del mismo MP descontaban el DOBLE
        # en producción. Sumamos su % y g_por_lote en una sola línea.
        _by_mid = {}
        for iv in items_validados:
            k = iv['material_id'].upper()
            if k in _by_mid:
                _by_mid[k]['porcentaje'] = round(_by_mid[k]['porcentaje'] + iv['porcentaje'], 4)
                _by_mid[k]['cantidad_g_por_lote'] = round(
                    _by_mid[k]['cantidad_g_por_lote'] + iv['cantidad_g_por_lote'], 2)
            else:
                _by_mid[k] = iv
        items_validados = list(_by_mid.values())
        # FIX 1-jun-2026 audit (P0-2) · validar suma de % · >101% es imposible (los
        # ingredientes no pueden superar el 100% del lote · error de captura). <95%
        # solo avisa (puede ser agua c.s.p. no listada).
        _sum_pct = round(sum(iv['porcentaje'] for iv in items_validados), 2)
        if _sum_pct > 101.0:
            return jsonify({
                'error': f'La suma de porcentajes es {_sum_pct}% (>100%). '
                         'Los ingredientes no pueden superar el 100% del lote · revisá los valores.',
                'suma_pct': _sum_pct,
            }), 400
        _warning_suma = (f'La suma de porcentajes es {_sum_pct}% (<100%). '
                         'Si el resto es agua/c.s.p. está bien; si falta un ingrediente, agregalo.'
                         ) if _sum_pct < 95.0 else None

        # GUARD nombre↔código · 1-jun-2026 (caso "N-acetil glucosamina" guardada
        # con MP00175 = Acetyl tetrapeptide-5 → stock cruzado → "Hay 0g" al
        # producir). El form tiene código y nombre como inputs independientes y el
        # backend solo validaba que el código existiera, NO que concordara con el
        # nombre. Acá rechazamos el mapeo cruzado salvo override explícito,
        # usando el MISMO motor que el detector (blueprints.formula_match · M1).
        if not data.get('forzar_mismatch'):
            try:
                from blueprints.formula_match import build_maestro_index, evaluar_item
            except ImportError:
                from api.blueprints.formula_match import build_maestro_index, evaluar_item
            try:
                _mrows = c.execute(
                    "SELECT codigo_mp, COALESCE(nombre_comercial,''), COALESCE(nombre_inci,'') "
                    "FROM maestro_mps WHERE activo=1").fetchall()
                _idx = build_maestro_index(_mrows)
                _mismatches = []
                for it in items_validados:
                    ev = evaluar_item(it['material_nombre'], it['material_id'], _idx)
                    _mej = ev.get('mejor_candidato') or {}
                    # Bloquear SOLO si el nombre choca con el código Y existe otro
                    # código que es un match FUERTE (≥80) y DISTINTO al asignado.
                    # Así no molestamos cuando el código asignado es lo mejor que hay
                    # (nombre abreviado/variante sin mejor candidato) · evita falsos
                    # positivos en nombres que no calzan léxicamente con el catálogo.
                    if (ev.get('problema') == 'mismatch_nombre'
                            and _mej.get('codigo')
                            and _mej.get('codigo') != it['material_id']
                            and (_mej.get('score') or 0) >= 80):
                        _mismatches.append({
                            'material_id': it['material_id'],
                            'material_nombre': it['material_nombre'],
                            'codigo_es_en_catalogo': ev.get('maestro_nombre') or '?',
                            'codigo_sugerido': _mej.get('codigo'),
                            'nombre_sugerido': _mej.get('nombre_comercial') or _mej.get('nombre_inci'),
                        })
            except Exception:
                _mismatches = []  # nunca bloquear por fallo del motor (degradación segura)
            if _mismatches:
                return jsonify({
                    'error': 'El nombre no coincide con el código en una o más líneas',
                    'detalle': ('El código pertenece a OTRA materia prima del catálogo. '
                                'Corregí el código de la línea marcada, o reenviá con '
                                '"forzar_mismatch" si estás 100% seguro.'),
                    'mismatches': _mismatches,
                    'forzar_mismatch_requerido': True,
                }), 409

        # Sprint Fórmulas PRO · 20-may-2026: versionar ANTES de pisar.
        # Si ya existe esa fórmula, archivamos versión actual en
        # formula_versiones. INVIMA puede pedir trazabilidad de cambios.
        prev_header = c.execute(
            "SELECT unidad_base_g, descripcion FROM formula_headers WHERE producto_nombre=?",
            (prod,),
        ).fetchone()
        if prev_header:
            prev_items = c.execute(
                "SELECT material_id, material_nombre, porcentaje, cantidad_g_por_lote "
                "FROM formula_items WHERE producto_nombre=?",
                (prod,),
            ).fetchall()
            if prev_items:
                try:
                    import json as _json
                    prev_items_json = _json.dumps([{
                        'material_id': r[0], 'material_nombre': r[1],
                        'porcentaje': r[2], 'cantidad_g_por_lote': r[3],
                    } for r in prev_items], ensure_ascii=False)
                    next_ver_row = c.execute(
                        "SELECT COALESCE(MAX(version),0)+1 FROM formula_versiones "
                        "WHERE producto_nombre=?", (prod,),
                    ).fetchone()
                    next_ver = next_ver_row[0] if next_ver_row else 1
                    motivo_change = (data.get('motivo_cambio') or '').strip()[:300]
                    c.execute(
                        """INSERT INTO formula_versiones
                             (producto_nombre, version, unidad_base_g,
                              descripcion, items_json, creado_por, motivo_cambio)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (prod, next_ver,
                         float(prev_header[0] or 1000),
                         prev_header[1] or '',
                         prev_items_json,
                         session.get('compras_user', '_'),
                         motivo_change or 'edición sin motivo'),
                    )
                except Exception as _e:
                    __import__('logging').getLogger('inventario').warning(
                        'formula_versiones snapshot fallo: %s', _e)

        try:
            c.execute(
                """INSERT OR REPLACE INTO formula_headers
                   (producto_nombre, unidad_base_g, lote_size_kg, descripcion, fecha_creacion)
                   VALUES (?,?,?,?,?)""",
                (prod, unidad_base_g, lote_size_kg, data.get('descripcion', ''),
                 datetime.now().isoformat())
            )
            c.execute('DELETE FROM formula_items WHERE producto_nombre=?', (prod,))
            for it in items_validados:
                c.execute(
                    """INSERT INTO formula_items
                       (producto_nombre, material_id, material_nombre,
                        porcentaje, cantidad_g_por_lote)
                       VALUES (?,?,?,?,?)""",
                    (prod, it['material_id'], it['material_nombre'],
                     it['porcentaje'], it['cantidad_g_por_lote'])
                )
            try:
                audit_log(c, usuario=session.get('compras_user',''),
                          accion='FORMULA_GUARDAR',
                          tabla='formula_headers', registro_id=prod,
                          despues={'unidad_base_g': unidad_base_g,
                                    'items_count': len(items_validados),
                                    'era_edicion': bool(prev_header)})
            except Exception:
                pass
            conn.commit()
            return jsonify({
                'message': f'Fórmula de {prod} guardada · {len(items_validados)} items',
                'producto': prod,
                'unidad_base_g': unidad_base_g,
                'lote_size_kg': lote_size_kg,
                'items_count': len(items_validados),
                'suma_pct': _sum_pct,
                'warning': _warning_suma,
            }), 201
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            return jsonify({
                'error': 'Falla guardando fórmula',
                'detalle': str(e)[:300],
            }), 500
    c.execute('SELECT producto_nombre, unidad_base_g, descripcion, fecha_creacion FROM formula_headers ORDER BY producto_nombre')
    headers = c.fetchall()
    formulas = []
    for h in headers:
        # INCI (13-jun · UI por INCI): JOIN al maestro por código (la identidad sigue
        # siendo material_id; nombre_inci es solo para mostrar).
        c.execute('SELECT fi.material_id, fi.material_nombre, fi.porcentaje, '
                  "COALESCE(mm.nombre_inci,'') FROM formula_items fi "
                  'LEFT JOIN maestro_mps mm ON mm.codigo_mp=fi.material_id '
                  'WHERE fi.producto_nombre=?', (h[0],))
        items = [{'material_id': r[0], 'material_nombre': r[1], 'porcentaje': r[2],
                  'nombre_inci': r[3]} for r in c.fetchall()]
        formulas.append({'producto_nombre': h[0], 'unidad_base_g': h[1], 'descripcion': h[2],
                         'fecha_creacion': h[3], 'items': items})
    return jsonify({'formulas': formulas})

@bp.route('/api/formulas/bases-stats', methods=['GET'])
def formulas_bases_stats():
    """Sprint Fórmulas PRO · Sebastián 20-may-2026: "las bases, todas
    tienen una base diferente, revisemos".

    Devuelve la distribución de `unidad_base_g` agrupada · permite
    detectar inconsistencias (Renova C 10 = 1000g vs otra = 500g).

    NOTA: `unidad_base_g` es solo el lote nominal · NO afecta los
    descuentos reales (que usan % × cantidad_kg pedida). Igual mostrar
    todas con la misma base ayuda a comparar fórmulas visualmente.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    rows = c.execute(
        """SELECT unidad_base_g,
                  COUNT(*) AS n,
                  GROUP_CONCAT(producto_nombre, ' | ') AS productos
           FROM formula_headers
           GROUP BY unidad_base_g
           ORDER BY n DESC, unidad_base_g""",
    ).fetchall()
    grupos = [{
        'unidad_base_g': float(r[0] or 0),
        'count': r[1],
        'productos': (r[2] or '').split(' | ') if r[2] else [],
    } for r in rows]
    total = sum(g['count'] for g in grupos)
    base_dominante = max(grupos, key=lambda g: g['count'])['unidad_base_g'] if grupos else None
    return jsonify({
        'grupos': grupos,
        'total_formulas': total,
        'base_dominante_g': base_dominante,
        'es_uniforme': len(grupos) <= 1,
        'mensaje': (
            f'Todas las fórmulas usan {base_dominante}g como base.'
            if len(grupos) <= 1
            else f'{len(grupos)} bases distintas detectadas · {base_dominante}g es la más común'
        ),
    })


@bp.route('/api/formulas/normalizar-base', methods=['POST'])
def formulas_normalizar_base():
    """Sprint Fórmulas PRO · admin · normaliza la base de todas las
    fórmulas (o un subset) a una unidad común.

    Body: {base_g: float, productos?: [str]}
    Si productos vacío → afecta TODAS. Si lista → solo esas.

    SAFE: los % NO se tocan · solo `unidad_base_g` y `cantidad_g_por_lote`
    (recalculado = pct × base_nueva / 100). Esto NO afecta los descuentos
    reales porque /api/produccion calcula con la cantidad_kg pedida × pct.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if (u or '').strip().lower() not in {x.lower() for x in ADMIN_USERS}:
        return jsonify({'error': 'solo admin'}), 403
    body = request.get_json(silent=True) or {}
    try:
        base_g = float(body.get('base_g') or 0)
    except (ValueError, TypeError):
        base_g = 0
    if base_g < 50 or base_g > 100000:
        return jsonify({'error': 'base_g debe ser 50-100000 (g)'}), 400
    productos_filtro = body.get('productos') or []
    conn = get_db(); c = conn.cursor()
    if productos_filtro:
        placeholders = ','.join(['?'] * len(productos_filtro))
        target = c.execute(
            f"SELECT producto_nombre FROM formula_headers "
            f"WHERE producto_nombre IN ({placeholders})",
            productos_filtro,
        ).fetchall()
    else:
        target = c.execute("SELECT producto_nombre FROM formula_headers").fetchall()
    actualizadas = []
    for (prod,) in target:
        try:
            c.execute(
                "UPDATE formula_headers SET unidad_base_g=?, "
                "lote_size_kg=? WHERE producto_nombre=?",
                (base_g, round(base_g / 1000.0, 3), prod),
            )
            # Recalcular cantidad_g_por_lote de cada item (pct × base / 100)
            c.execute(
                "UPDATE formula_items "
                "SET cantidad_g_por_lote = ROUND(porcentaje * ? / 100.0, 2) "
                "WHERE producto_nombre = ?",
                (base_g, prod),
            )
            actualizadas.append(prod)
        except Exception:
            continue
    try:
        audit_log(c, usuario=u, accion='FORMULAS_NORMALIZAR_BASE',
                  tabla='formula_headers', registro_id='_BULK_',
                  despues={'base_g': base_g,
                            'actualizadas': len(actualizadas)})
    except Exception:
        pass
    conn.commit()
    return jsonify({
        'ok': True,
        'base_g': base_g,
        'actualizadas_count': len(actualizadas),
        'productos': actualizadas[:50],
        'mensaje': f'{len(actualizadas)} fórmulas normalizadas a base {base_g}g',
    })


@bp.route('/api/formulas/<path:producto_nombre>/versiones', methods=['GET'])
def formulas_versiones(producto_nombre):
    """Sprint Fórmulas PRO · historial de versiones de una fórmula."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    try:
        rows = c.execute(
            """SELECT id, version, unidad_base_g, descripcion, items_json,
                      creado_at_utc, creado_por, motivo_cambio
               FROM formula_versiones
               WHERE producto_nombre=?
               ORDER BY version DESC LIMIT 50""",
            (producto_nombre,),
        ).fetchall()
    except Exception:
        rows = []
    import json as _json
    items = []
    for r in rows:
        try:
            its = _json.loads(r[4]) if r[4] else []
        except Exception:
            its = []
        items.append({
            'id': r[0], 'version': r[1],
            'unidad_base_g': float(r[2] or 0),
            'descripcion': r[3] or '',
            'items_count': len(its), 'items': its,
            'creado_at_utc': r[5], 'creado_por': r[6] or '',
            'motivo_cambio': r[7] or '',
        })
    return jsonify({'producto': producto_nombre, 'versiones': items,
                    'total': len(items)})


@bp.route('/api/formulas/duplicar', methods=['POST'])
def formulas_duplicar():
    """Sprint Fórmulas PRO · duplica una fórmula con nombre nuevo.

    Body: {producto_origen, producto_nuevo}
    Útil para crear variantes (Renova C 10 → Renova C 5).
    """
    u, err, code = _require_session()
    if err:
        return err, code
    body = request.get_json(silent=True) or {}
    origen = (body.get('producto_origen') or '').strip()
    nuevo = (body.get('producto_nuevo') or '').strip()
    if not origen or not nuevo:
        return jsonify({'error': 'producto_origen y producto_nuevo requeridos'}), 400
    conn = get_db(); c = conn.cursor()
    h_orig = c.execute(
        "SELECT unidad_base_g, lote_size_kg, descripcion FROM formula_headers "
        "WHERE producto_nombre=?", (origen,),
    ).fetchone()
    if not h_orig:
        return jsonify({'error': f'fórmula origen "{origen}" no existe'}), 404
    if c.execute(
        "SELECT 1 FROM formula_headers WHERE LOWER(producto_nombre)=LOWER(?)",
        (nuevo,),
    ).fetchone():
        return jsonify({'error': f'fórmula "{nuevo}" ya existe'}), 409
    c.execute(
        "INSERT INTO formula_headers (producto_nombre, unidad_base_g, lote_size_kg, descripcion, fecha_creacion) VALUES (?,?,?,?,?)",
        (nuevo, h_orig[0], h_orig[1],
         (h_orig[2] or '') + ' · duplicado de ' + origen,
         datetime.now().isoformat()),
    )
    items_orig = c.execute(
        "SELECT material_id, material_nombre, porcentaje, cantidad_g_por_lote "
        "FROM formula_items WHERE producto_nombre=?", (origen,),
    ).fetchall()
    for it in items_orig:
        c.execute(
            "INSERT INTO formula_items (producto_nombre, material_id, material_nombre, porcentaje, cantidad_g_por_lote) VALUES (?,?,?,?,?)",
            (nuevo, it[0], it[1], it[2], it[3]),
        )
    try:
        audit_log(c, usuario=u, accion='FORMULA_DUPLICAR',
                  tabla='formula_headers', registro_id=nuevo,
                  despues={'origen': origen, 'items_count': len(items_orig)})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'producto': nuevo,
                    'mensaje': f'Fórmula duplicada de {origen}',
                    'items_count': len(items_orig)})


@bp.route('/api/formulas/import-excel', methods=['POST'])
def formulas_import_excel():
    """Sprint Fórmulas PRO · 20-may-2026 · IMPORT EXCEL.

    Pedido directo de Sebastián: "Alejandro crea las fórmulas y las envía
    en Excel, entonces que se pueda cargar y la app la vuelva fórmula".

    Body: archivo XLSX/CSV en el campo `file` (multipart) o `contenido_b64`
    en JSON.

    Formato esperado · 1 fila por ingrediente:
      producto | unidad_base_g | descripcion | codigo_mp | nombre_mp | porcentaje

    Hojas se ignoran · solo se lee la primera. Headers se buscan
    case-insensitive con variantes ("producto"/"product"/"nombre_producto",
    "% en fórmula"/"porcentaje"/"%"/"pct", etc.). dry_run=true devuelve
    preview sin guardar.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'solo admin puede importar fórmulas masivo'}), 403
    dry_run = (request.args.get('dry_run') or '').strip() in ('1','true','yes')

    # Leer file desde multipart o b64
    raw_bytes = None
    filename = ''
    if request.files and 'file' in request.files:
        fp = request.files['file']
        filename = fp.filename or 'upload.xlsx'
        raw_bytes = fp.read()
    else:
        body = request.get_json(silent=True) or {}
        b64 = body.get('contenido_b64') or ''
        if b64:
            import base64 as _b64
            try:
                raw_bytes = _b64.b64decode(b64)
                filename = body.get('filename') or 'upload.xlsx'
            except Exception:
                return jsonify({'error': 'base64 inválido'}), 400
    if not raw_bytes:
        return jsonify({'error': 'archivo requerido (file= o contenido_b64)'}), 400
    if len(raw_bytes) > 10 * 1024 * 1024:
        return jsonify({'error': 'archivo muy grande (max 10MB)'}), 413

    # Parser · soporta XLSX (openpyxl) y CSV/TSV
    is_xlsx = filename.lower().endswith('.xlsx')
    rows_raw = []
    try:
        if is_xlsx:
            try:
                import openpyxl  # type: ignore
            except ImportError:
                return jsonify({
                    'error': 'openpyxl no disponible · subí el archivo como CSV',
                    'codigo': 'NO_OPENPYXL',
                }), 503
            import io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows_raw.append([
                    (str(c).strip() if c is not None else '') for c in row
                ])
        else:
            text = raw_bytes.decode('utf-8-sig', errors='replace')
            import csv as _csv, io as _io
            sniffer = _csv.Sniffer()
            try:
                dialect = sniffer.sniff(text[:2000], delimiters=',;\t|')
            except Exception:
                dialect = _csv.excel
            reader = _csv.reader(_io.StringIO(text), dialect)
            rows_raw = [list(r) for r in reader]
    except Exception as e:
        return jsonify({'error': f'parser falló: {e}'}), 400

    if len(rows_raw) < 2:
        return jsonify({'error': 'archivo vacío o sin headers'}), 400

    # Detectar columnas
    headers = [str(h).strip().lower() for h in rows_raw[0]]
    def _find(*keys):
        for i, h in enumerate(headers):
            for k in keys:
                if k in h:
                    return i
        return -1
    col_prod = _find('producto', 'product', 'nombre_producto', 'producto_nombre')
    col_base = _find('unidad_base', 'base', 'base_g', 'lote_g')
    col_desc = _find('descripcion', 'description')
    col_mp_cod = _find('codigo_mp', 'cod_mp', 'codigo', 'code', 'sku')
    col_mp_nom = _find('nombre_mp', 'nombre_material', 'material', 'ingrediente', 'mp_nombre')
    col_pct = _find('porcentaje', '% en', '%', 'pct', 'percent')
    missing = []
    if col_prod == -1: missing.append('producto')
    if col_mp_cod == -1: missing.append('codigo_mp')
    if col_pct == -1: missing.append('porcentaje')
    if missing:
        return jsonify({
            'error': f'faltan columnas requeridas: {missing}',
            'headers_detectados': headers,
            'hint': 'Usá columnas: producto, codigo_mp, porcentaje (opcional: nombre_mp, unidad_base_g, descripcion)',
        }), 400

    # Agrupar por producto
    formulas_por_prod = {}
    errores_filas = []
    for idx, row in enumerate(rows_raw[1:], start=2):
        if not row or all(not c for c in row):
            continue
        try:
            prod = (row[col_prod] if col_prod < len(row) else '').strip()
            if not prod:
                continue
            mid = (row[col_mp_cod] if col_mp_cod < len(row) else '').strip().upper()
            if not mid:
                errores_filas.append({'fila': idx, 'razon': 'codigo_mp vacío'})
                continue
            try:
                pct = float(str(row[col_pct] if col_pct < len(row) else '0').replace(',','.'))
            except (ValueError, TypeError):
                errores_filas.append({'fila': idx, 'razon': f'porcentaje inválido: {row[col_pct]}'})
                continue
            if pct <= 0:
                continue
            nom = (row[col_mp_nom] if col_mp_nom >= 0 and col_mp_nom < len(row) else '').strip()
            base = 1000.0
            if col_base >= 0 and col_base < len(row):
                try:
                    base = float(str(row[col_base]).replace(',','.') or 1000)
                except (ValueError, TypeError):
                    pass
            desc = (row[col_desc] if col_desc >= 0 and col_desc < len(row) else '').strip()
            formulas_por_prod.setdefault(prod, {
                'producto_nombre': prod,
                'unidad_base_g': base,
                'descripcion': desc,
                'items': [],
            })
            formulas_por_prod[prod]['items'].append({
                'material_id': mid,
                'material_nombre': nom or mid,
                'porcentaje': pct,
            })
        except Exception as _e:
            errores_filas.append({'fila': idx, 'razon': f'parse error: {_e}'})

    # Validar contra maestro_mps
    conn = get_db(); c = conn.cursor()
    plan = []
    for prod_nom, f in formulas_por_prod.items():
        total_pct = sum(it['porcentaje'] for it in f['items'])
        mps_faltantes = []
        for it in f['items']:
            row = c.execute(
                "SELECT nombre_comercial FROM maestro_mps WHERE codigo_mp=? AND COALESCE(activo,1)=1",
                (it['material_id'],),
            ).fetchone()
            if not row:
                mps_faltantes.append(it['material_id'])
            elif not it['material_nombre']:
                it['material_nombre'] = row[0]
        existe = c.execute(
            "SELECT 1 FROM formula_headers WHERE LOWER(producto_nombre)=LOWER(?)",
            (prod_nom,),
        ).fetchone()
        plan.append({
            'producto': prod_nom,
            'unidad_base_g': f['unidad_base_g'],
            'items_count': len(f['items']),
            'total_pct': round(total_pct, 2),
            'pct_ok': abs(total_pct - 100) < 1,
            'ya_existe': bool(existe),
            'mps_faltantes': mps_faltantes,
            'descripcion': f['descripcion'],
        })

    if dry_run:
        return jsonify({
            'dry_run': True,
            'archivo': filename,
            'formulas_detectadas': len(plan),
            'plan': plan,
            'errores_filas': errores_filas[:50],
            'headers_detectados': headers,
        })

    # APPLY
    aplicadas = []
    rechazadas = []
    for prod_nom, f in formulas_por_prod.items():
        # Validar MPs antes de tocar BD
        mps_faltantes = [it['material_id'] for it in f['items']
                         if not c.execute(
                             "SELECT 1 FROM maestro_mps WHERE codigo_mp=? AND COALESCE(activo,1)=1",
                             (it['material_id'],),
                         ).fetchone()]
        if mps_faltantes:
            rechazadas.append({
                'producto': prod_nom, 'razon': 'MPs no existen',
                'mps': mps_faltantes,
            })
            continue
        try:
            unidad_base_g = float(f['unidad_base_g'])
            # Reusa la lógica del POST /api/formulas via test_request_context
            # Simplificado: insertar directo
            c.execute(
                """INSERT OR REPLACE INTO formula_headers
                   (producto_nombre, unidad_base_g, lote_size_kg, descripcion, fecha_creacion)
                   VALUES (?,?,?,?,?)""",
                (prod_nom, unidad_base_g, round(unidad_base_g / 1000.0, 3),
                 f['descripcion'], datetime.now().isoformat()),
            )
            c.execute('DELETE FROM formula_items WHERE producto_nombre=?', (prod_nom,))
            for it in f['items']:
                gpl = round((it['porcentaje'] / 100.0) * unidad_base_g, 2)
                c.execute(
                    """INSERT INTO formula_items
                       (producto_nombre, material_id, material_nombre,
                        porcentaje, cantidad_g_por_lote)
                       VALUES (?,?,?,?,?)""",
                    (prod_nom, it['material_id'], it['material_nombre'],
                     it['porcentaje'], gpl),
                )
            aplicadas.append({
                'producto': prod_nom, 'items_count': len(f['items']),
            })
        except Exception as e:
            rechazadas.append({'producto': prod_nom, 'razon': str(e)[:200]})
    try:
        audit_log(c, usuario=u, accion='FORMULAS_IMPORT_EXCEL',
                  tabla='formula_headers', registro_id='_BULK_',
                  despues={'archivo': filename,
                           'aplicadas': len(aplicadas),
                           'rechazadas': len(rechazadas)})
    except Exception:
        pass
    conn.commit()
    return jsonify({
        'ok': True, 'archivo': filename,
        'aplicadas': aplicadas,
        'rechazadas': rechazadas,
        'errores_filas': errores_filas[:50],
        'mensaje': f'{len(aplicadas)} fórmulas importadas · {len(rechazadas)} rechazadas',
    })


@bp.route('/api/formulas/export-excel', methods=['GET'])
def formulas_export_excel():
    """Sprint Fórmulas PRO · export para que Alejandro edite en Excel y
    luego importe. Devuelve XLS HTML (mismo método que otros exports)."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    rows = c.execute(
        """SELECT h.producto_nombre, h.unidad_base_g, COALESCE(h.descripcion,''),
                  i.material_id, COALESCE(i.material_nombre,''),
                  i.porcentaje, COALESCE(i.cantidad_g_por_lote, 0)
           FROM formula_headers h
           LEFT JOIN formula_items i ON i.producto_nombre = h.producto_nombre
           ORDER BY h.producto_nombre, i.material_id""",
    ).fetchall()
    def _esc(s):
        return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
    body = ''.join(
        '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'
        .format(_esc(r[0]), r[1] or '', _esc(r[2]), _esc(r[3]),
                _esc(r[4]), r[5] or '', r[6] or '')
        for r in rows
    )
    html = (
        '<!DOCTYPE html><html><head><meta charset="utf-8"></head><body>'
        '<table border="1"><thead><tr>'
        '<th>producto</th><th>unidad_base_g</th><th>descripcion</th>'
        '<th>codigo_mp</th><th>nombre_mp</th><th>porcentaje</th>'
        '<th>cantidad_g_por_lote</th></tr></thead><tbody>'
        + body + '</tbody></table></body></html>'
    )
    from datetime import date as _d
    fname = f'formulas_eos_{_d.today().isoformat()}.xls'
    from flask import Response as _Resp
    resp = _Resp('﻿' + html, mimetype='application/vnd.ms-excel; charset=utf-8')
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@bp.route('/api/formulas/<path:producto_nombre>/uso', methods=['GET'])
def formulas_uso(producto_nombre):
    """Sprint Fórmulas PRO · cuántos lotes han usado esta fórmula."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    try:
        row = c.execute(
            """SELECT COUNT(*),
                      MAX(fecha_programada) as ultima_prog,
                      MAX(fin_real_at) as ultimo_real,
                      COALESCE(SUM(COALESCE(kg_real,cantidad_kg,0)),0) as kg_total
               FROM produccion_programada
               WHERE LOWER(producto) = LOWER(?)""",
            (producto_nombre,),
        ).fetchone()
    except Exception:
        row = (0, None, None, 0)
    return jsonify({
        'producto': producto_nombre,
        'lotes_total': row[0] or 0,
        'ultima_programada': row[1] or '',
        'ultima_terminada': row[2] or '',
        'kg_producido_total': float(row[3] or 0),
    })


@bp.route('/api/formulas/<producto_nombre>', methods=['DELETE'])
def del_formula(producto_nombre):
    # RBAC · borrar una fórmula es destructivo sobre datos regulados (INVIMA).
    # Solo ADMIN o CALIDAD, igual que patch_codigo_pt (mismo dato).
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS and user not in CALIDAD_USERS:
        return jsonify({'error': 'requiere admin o calidad'}), 403
    conn = get_db()
    c = conn.cursor()
    # Snapshot antes del borrado para trazabilidad (audit_log obligatorio).
    n_items = c.execute('SELECT COUNT(*) FROM formula_items WHERE producto_nombre=?',
                        (producto_nombre,)).fetchone()[0]
    hdr = c.execute('SELECT COUNT(*) FROM formula_headers WHERE producto_nombre=?',
                    (producto_nombre,)).fetchone()[0]
    if not hdr and not n_items:
        return jsonify({'error': 'fórmula no encontrada'}), 404
    c.execute('DELETE FROM formula_items WHERE producto_nombre=?', (producto_nombre,))
    c.execute('DELETE FROM formula_headers WHERE producto_nombre=?', (producto_nombre,))
    audit_log(c, usuario=user, accion='ELIMINAR_FORMULA',
              tabla='formula_headers', registro_id=producto_nombre,
              antes={'producto_nombre': producto_nombre, 'items': n_items})
    conn.commit()
    return jsonify({'message': f'Formula {producto_nombre} eliminada'})


@bp.route('/api/formulas/<path:producto_nombre>/codigo-pt', methods=['PATCH'])
def patch_codigo_pt(producto_nombre):
    """Asigna o limpia el codigo_pt (MyBatch-compat) de un producto.

    Solo ADMIN o CALIDAD pueden modificar. UNIQUE constraint del índice
    parcial bloquea duplicados (mig 117). Body: {codigo_pt: "BB001"} o
    {codigo_pt: null} para limpiar.
    """
    user = session.get('compras_user', '')
    if user not in ADMIN_USERS and user not in CALIDAD_USERS:
        return jsonify({'error': 'requiere admin o calidad'}), 403

    body = request.get_json(silent=True) or {}
    raw = body.get('codigo_pt')
    if raw is None or raw == '':
        nuevo = None
    else:
        nuevo = str(raw).strip().upper()
        if not nuevo or len(nuevo) > 20:
            return jsonify({'error': 'codigo_pt vacío o > 20 chars'}), 400

    conn = get_db()
    c = conn.cursor()
    row = c.execute(
        'SELECT codigo_pt FROM formula_headers WHERE producto_nombre = ?',
        (producto_nombre,),
    ).fetchone()
    if not row:
        return jsonify({'error': 'producto no encontrado'}), 404

    anterior = row[0]
    try:
        c.execute(
            'UPDATE formula_headers SET codigo_pt = ? WHERE producto_nombre = ?',
            (nuevo, producto_nombre),
        )
        conn.commit()
    except sqlite3.IntegrityError as e:
        # UNIQUE índice parcial bloqueó duplicado
        if 'UNIQUE' in str(e).upper():
            return jsonify({'error': f'codigo_pt "{nuevo}" ya asignado a otro producto'}), 409
        raise

    audit_log(c, usuario=user, accion='SET_CODIGO_PT',
              tabla='formula_headers', registro_id=producto_nombre,
              antes={'codigo_pt': anterior},
              despues={'codigo_pt': nuevo})
    conn.commit()
    return jsonify({'ok': True, 'producto_nombre': producto_nombre,
                     'codigo_pt': nuevo, 'anterior': anterior})


@bp.route('/api/formulas/<path:producto_nombre>/imagen', methods=['POST', 'GET', 'DELETE'])
def producto_imagen(producto_nombre):
    """Gestiona la URL de imagen de un producto (para mostrar en checklist).

    POST  body: {imagen_url: "https://..."}  → guarda
    GET                                       → devuelve URL actual
    DELETE                                    → limpia URL

    Sebastian (28-abr-2026): Permite que el modal del checklist
    Pre-Produccion muestre la foto del producto (tomada de animuslb.com /
    Shopify) para que el equipo reconozca cual es visualmente.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    if request.method == 'GET':
        r = c.execute(
            "SELECT COALESCE(imagen_url,'') FROM formula_headers WHERE producto_nombre=?",
            (producto_nombre,)
        ).fetchone()
        return jsonify({'producto': producto_nombre, 'imagen_url': r[0] if r else ''})

    if request.method == 'DELETE':
        c.execute(
            "UPDATE formula_headers SET imagen_url='', imagen_actualizada_at=datetime('now', '-5 hours') "
            "WHERE producto_nombre=?",
            (producto_nombre,)
        )
        conn.commit()
        return jsonify({'ok': True, 'producto': producto_nombre, 'imagen_url': ''})

    # POST
    d = request.json or {}
    url = (d.get('imagen_url') or '').strip()
    if url and not url.startswith(('http://', 'https://', '/static/')):
        return jsonify({'error': 'URL invalida (debe empezar con http(s):// o /static/)'}), 400
    cur = c.execute(
        "UPDATE formula_headers SET imagen_url=?, imagen_actualizada_at=datetime('now', '-5 hours') "
        "WHERE producto_nombre=?",
        (url, producto_nombre)
    )
    if cur.rowcount == 0:
        return jsonify({'error': f'Producto {producto_nombre} no existe en formula_headers'}), 404
    conn.commit()
    return jsonify({'ok': True, 'producto': producto_nombre, 'imagen_url': url})


def _shopify_creds(conn):
    """Lee shopify_token y shopify_shop de animus_config. Retorna (token,shop)
    o (None,None) si no estan configurados.
    """
    try:
        rt = conn.execute("SELECT valor FROM animus_config WHERE clave='shopify_token'").fetchone()
        rs = conn.execute("SELECT valor FROM animus_config WHERE clave='shopify_shop'").fetchone()
    except Exception:
        return None, None
    if not rt or not rs:
        return None, None
    return rt[0], rs[0]


def _normalizar_nombre(s):
    """Normaliza para fuzzy match: lowercase, sin acentos, sin %  +%, etc."""
    import unicodedata
    s = (s or '').strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')  # sin acentos
    # quitar puntuacion + extra spaces
    import re as _re
    s = _re.sub(r'[^a-z0-9\s]', ' ', s)
    s = _re.sub(r'\s+', ' ', s).strip()
    return s


_SHOPIFY_CATALOG_CACHE = {'data': None, 'fetched_at': 0}


def _shopify_get_all_products(token, shop, timeout=20):
    """Lista todos los productos de Shopify (con cache 5 min)."""
    import time as _t, urllib.request, json as _json
    now = _t.time()
    if _SHOPIFY_CATALOG_CACHE['data'] and (now - _SHOPIFY_CATALOG_CACHE['fetched_at']) < 300:
        return _SHOPIFY_CATALOG_CACHE['data']
    products = []
    page_url = f"https://{shop}/admin/api/2024-01/products.json?limit=250"
    try:
        for _ in range(10):  # max 10 paginas (~2500 productos)
            req = urllib.request.Request(
                page_url,
                headers={'X-Shopify-Access-Token': token}
            )
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = _json.loads(resp.read().decode('utf-8'))
                link_header = resp.headers.get('Link', '')
            page_products = data.get('products', [])
            products.extend(page_products)
            # Buscar next page en Link header
            if 'rel="next"' not in link_header:
                break
            import re as _re
            m = _re.search(r'<([^>]+)>;\s*rel="next"', link_header)
            if not m:
                break
            page_url = m.group(1)
    except Exception:
        pass
    _SHOPIFY_CATALOG_CACHE['data'] = products
    _SHOPIFY_CATALOG_CACHE['fetched_at'] = now
    return products


def _shopify_sync_producto(conn, producto_nombre, token=None, shop=None, timeout=10):
    """Sync de UN producto desde Shopify a formula_headers.

    Estrategia de match (en orden):
      1. ?title=<nombre> exacto
      2. Match contra catalogo completo por nombre normalizado (sin acentos)
      3. Substring contains en titles del catalogo

    Devuelve dict {ok, ...} o {error, ...}.
    """
    if not token or not shop:
        token, shop = _shopify_creds(conn)
    if not token or not shop:
        return {'error': 'Shopify no configurado'}
    import urllib.request, urllib.parse, json as _json, re as _re
    base = f"https://{shop}/admin/api/2024-01/products.json"

    # Estrategia 1: title exacto
    qs = urllib.parse.urlencode({'title': producto_nombre, 'limit': 5})
    req = urllib.request.Request(
        f"{base}?{qs}",
        headers={'X-Shopify-Access-Token': token}
    )
    products = []
    try:
        resp = urllib.request.urlopen(req, timeout=timeout)
        data = _json.loads(resp.read().decode('utf-8'))
        products = data.get('products', []) or []
    except Exception:
        pass

    # Estrategia 0-3: si no hubo match exacto por título, traer catálogo completo
    matched_strategy = 'exact'
    if not products:
        all_products = _shopify_get_all_products(token, shop, timeout=20)
        # 0. (MÁS CONFIABLE) match por SKU mapeado del producto · el nombre en
        # Shopify puede diferir del de la fórmula, pero el SKU es la llave exacta.
        try:
            _skus = {(r[0] or '').strip().upper() for r in conn.execute(
                "SELECT sku FROM sku_producto_map WHERE COALESCE(activo,1)=1 "
                "AND UPPER(TRIM(producto_nombre))=UPPER(TRIM(?))", (producto_nombre,)).fetchall()}
            _skus.discard('')
        except Exception:
            _skus = set()
        if _skus:
            for p in all_products:
                if any((v.get('sku') or '').strip().upper() in _skus for v in (p.get('variants') or [])):
                    products = [p]
                    matched_strategy = 'sku'
                    break
        target_norm = _normalizar_nombre(producto_nombre)
        # 2. Normalizado igual
        if not products:
            for p in all_products:
                if _normalizar_nombre(p.get('title', '')) == target_norm:
                    products = [p]
                    matched_strategy = 'normalized_eq'
                    break
        # 3. Substring contains (Jaccard ≥ 0.5)
        if not products:
            target_words = set(target_norm.split())
            best = None; best_score = 0
            for p in all_products:
                cand_norm = _normalizar_nombre(p.get('title', ''))
                cand_words = set(cand_norm.split())
                if not cand_words:
                    continue
                inter = target_words & cand_words
                if not inter:
                    continue
                score = len(inter) / max(len(target_words), len(cand_words))
                if score > best_score and score >= 0.5:
                    best = p; best_score = score
            if best:
                products = [best]
                matched_strategy = f'fuzzy_{best_score:.2f}'

    if not products:
        return {'error': f'No encontrado en Shopify: "{producto_nombre}"', 'not_found': True}

    p0 = products[0]
    imgs = p0.get('images', []) or []
    imagen_principal = imgs[0].get('src', '') if imgs else ''
    imagenes_extra = [{'src': im.get('src', ''), 'alt': im.get('alt', ''), 'position': im.get('position', 0)}
                       for im in imgs]
    variants = p0.get('variants', []) or []
    var0 = variants[0] if variants else {}
    sku_principal = var0.get('sku', '') or ''
    precio_venta = float(var0.get('price') or 0)
    peso_g = float(var0.get('grams') or 0)
    body_html = p0.get('body_html', '') or ''
    descripcion_plain = _re.sub(r'<[^>]+>', ' ', body_html)
    descripcion_plain = _re.sub(r'\s+', ' ', descripcion_plain).strip()[:500]

    cur = conn.execute("""
        UPDATE formula_headers SET
          imagen_url=?, imagen_actualizada_at=datetime('now', '-5 hours'),
          shopify_id=?, shopify_handle=?,
          descripcion_html=?, descripcion_plain=?,
          sku_principal=?, precio_venta=?, peso_g=?,
          imagenes_extra_json=?, shopify_synced_at=datetime('now', '-5 hours')
        WHERE producto_nombre=?
    """, (
        imagen_principal, str(p0.get('id') or ''), p0.get('handle') or '',
        body_html, descripcion_plain,
        sku_principal, precio_venta, peso_g,
        _json.dumps(imagenes_extra), producto_nombre,
    ))
    if cur.rowcount == 0:
        return {'error': f'Producto {producto_nombre} no existe en formula_headers'}
    conn.commit()
    return {
        'ok': True, 'producto': producto_nombre,
        'imagen_url': imagen_principal, 'imagenes_count': len(imagenes_extra),
        'sku': sku_principal, 'precio': precio_venta, 'peso_g': peso_g,
        'shopify_product_id': p0.get('id'), 'shopify_handle': p0.get('handle'),
    }


def _sync_shopify_pendientes_background(max_edad_horas=24, max_productos=50):
    """Sincroniza en BACKGROUND todos los productos cuyo shopify_synced_at sea
    NULL o mas viejo que max_edad_horas. No bloquea el request HTTP que lo
    invoco — corre en thread separado.

    Idempotente: si ya esta corriendo en otro thread, retorna sin hacer nada.
    """
    import threading
    if getattr(_sync_shopify_pendientes_background, '_running', False):
        return  # ya hay otro thread sincronizando
    _sync_shopify_pendientes_background._running = True

    def _worker():
        try:
            from config import DB_PATH
            import sqlite3
            local_conn = db_connect(timeout=30)
            try:
                token, shop = _shopify_creds(local_conn)
                if not token or not shop:
                    return
                # SEC-FIX · 21-may-2026 · param `?` en vez de f-string (anti-SQLi defensa profunda)
                # Calcular cutoff en Python (compatible PG · evita date multi-arg)
                from datetime import datetime as _dtcut, timedelta as _tdcut
                cutoff_str = (_dtcut.now() - _tdcut(hours=5 + int(max_edad_horas))).isoformat()
                rows = local_conn.execute("""
                    SELECT producto_nombre FROM formula_headers
                    WHERE COALESCE(shopify_synced_at,'') = ''
                       OR shopify_synced_at < ?
                    LIMIT ?
                """, (cutoff_str, max_productos)).fetchall()
                for (prod,) in rows:
                    try:
                        _shopify_sync_producto(local_conn, prod, token, shop, timeout=8)
                    except Exception:
                        continue
                    # Pequeño throttle para no saturar Shopify (rate limit ~2 req/s)
                    import time as _t; _t.sleep(0.6)
            finally:
                local_conn.close()
        finally:
            _sync_shopify_pendientes_background._running = False

    t = threading.Thread(target=_worker, daemon=True)
    t.start()


@bp.route('/api/formulas/<path:producto_nombre>/imagen-shopify-sync', methods=['POST'])
def producto_imagen_shopify_sync(producto_nombre):
    """Sync de UN producto desde Shopify (manual via boton del modal).

    Usa el helper compartido _shopify_sync_producto.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    res = _shopify_sync_producto(conn, producto_nombre, timeout=15)
    if res.get('error'):
        status = 404 if res.get('not_found') else (400 if 'no configurado' in res['error'] else 502)
        return jsonify(res), status
    return jsonify(res)


# ─── Proxy de imagenes Shopify ────────────────────────────────────────
# Las imagenes guardadas vienen de cdn.shopify.com. Algunos browsers / setups
# bloquean hotlink directo (CORS / referer policy / lazy network). Servimos
# las imagenes via proxy desde nuestro propio dominio para evitar problemas.
# Cache simple en memoria: dict {url -> (content_bytes, mime, fetched_at)}
_IMG_CACHE = {}
_IMG_CACHE_MAX = 200      # max entradas (LRU implicito)
_IMG_CACHE_TTL = 3600 * 6  # 6 horas


def _fetch_imagen_remota(url, timeout=10):
    """Descarga imagen de URL remota con cache simple. Devuelve (bytes, mime)
    o (None, None) si falla.
    """
    import time as _t
    now = _t.time()
    hit = _IMG_CACHE.get(url)
    if hit and (now - hit[2]) < _IMG_CACHE_TTL:
        return hit[0], hit[1]
    import urllib.request
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'EOS-Holding/1.0 (image-proxy)',
            'Accept': 'image/*,*/*;q=0.8',
        })
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
            mime = resp.headers.get('Content-Type', 'image/jpeg').split(';')[0].strip()
    except Exception:
        return None, None
    # LRU eviction simple
    if len(_IMG_CACHE) >= _IMG_CACHE_MAX:
        oldest = min(_IMG_CACHE.items(), key=lambda kv: kv[1][2])
        _IMG_CACHE.pop(oldest[0], None)
    _IMG_CACHE[url] = (data, mime, now)
    return data, mime


@bp.route('/api/imagen-producto/<path:producto_nombre>', methods=['GET'])
def imagen_producto_proxy(producto_nombre):
    """Sirve la imagen principal del producto via proxy.

    Resuelve el problema de hotlink/CORS al mostrar imagenes de Shopify CDN
    desde nuestro dominio. El navegador hace request a /api/imagen-producto/
    <nombre> en vez de cdn.shopify.com.

    Querystring opcional: ?idx=N para servir la imagen N de imagenes_extra.
    """
    if 'compras_user' not in session:
        return ('Unauthorized', 401)
    conn = get_db()
    row = conn.execute(
        "SELECT COALESCE(imagen_url,''), COALESCE(imagenes_extra_json,'[]') "
        "FROM formula_headers WHERE producto_nombre=?",
        (producto_nombre,)
    ).fetchone()
    if not row:
        return ('Not Found', 404)
    imagen_url, imgs_json = row[0], row[1]

    idx = request.args.get('idx')
    if idx is not None:
        import json as _json
        try:
            imgs = _json.loads(imgs_json or '[]')
            n = int(idx)
            if 0 <= n < len(imgs):
                imagen_url = imgs[n].get('src', '') or imagen_url
        except Exception:
            pass

    if not imagen_url:
        return ('No image', 404)
    data, mime = _fetch_imagen_remota(imagen_url, timeout=10)
    if data is None:
        return ('Upstream fetch failed', 502)
    resp = Response(data, mimetype=mime)
    # Cache navegador: 1h. Si re-syncamos en BD, el bumpeo de URL cambia el cache key
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp


@bp.route('/api/formulas/catalogo', methods=['GET'])
def formulas_catalogo():
    """Lista todos los productos de formula_headers con su estado de sync.

    Útil para ver de un vistazo cuáles tienen foto y cuáles no, y forzar
    sync de los pendientes.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    rows = conn.execute("""
        SELECT producto_nombre,
               COALESCE(imagen_url,''),
               COALESCE(sku_principal,''),
               COALESCE(precio_venta,0),
               COALESCE(shopify_handle,''),
               COALESCE(shopify_synced_at,'')
        FROM formula_headers
        ORDER BY
          CASE WHEN COALESCE(imagen_url,'')='' THEN 0 ELSE 1 END,
          producto_nombre
    """).fetchall()
    productos = []
    for r in rows:
        productos.append({
            'nombre': r[0],
            'imagen_url':       r[1],
            'sku':              r[2],
            'precio':           float(r[3] or 0),
            'shopify_handle':   r[4],
            'sincronizado':     bool(r[5]),
            'synced_at':        r[5],
            'tiene_foto':       bool(r[1]),
        })
    con_foto = sum(1 for p in productos if p['tiene_foto'])
    sin_foto = len(productos) - con_foto
    return jsonify({
        'productos': productos,
        'total': len(productos),
        'con_foto': con_foto,
        'sin_foto': sin_foto,
    })


@bp.route('/api/formulas/sync-shopify-blocking', methods=['POST'])
def sync_shopify_blocking():
    """Sincroniza TODOS los productos pendientes de forma SINCRONA y devuelve
    el resultado por producto. Usar para 'sync ahora' con feedback inmediato.

    Querystring: ?force=1 para re-sincronizar incluso los ya sincronizados.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    token, shop = _shopify_creds(conn)
    if not token or not shop:
        return jsonify({'error': 'Shopify no configurado en animus_config'}), 400
    force = request.args.get('force', '0') == '1'

    where = "" if force else "WHERE COALESCE(shopify_synced_at,'') = '' OR shopify_synced_at < datetime('now', '-5 hours', '-24 hours')"
    rows = conn.execute(
        f"SELECT producto_nombre FROM formula_headers {where} LIMIT 200"
    ).fetchall()
    resultados = {'ok': [], 'no_encontrados': [], 'errores': []}
    import time as _t
    for (prod,) in rows:
        try:
            res = _shopify_sync_producto(conn, prod, token, shop, timeout=12)
        except Exception as e:
            resultados['errores'].append({'producto': prod, 'error': str(e)})
            continue
        if res.get('ok'):
            resultados['ok'].append({
                'producto': prod,
                'imagen': bool(res.get('imagen_url')),
                'sku': res.get('sku', ''),
            })
        elif res.get('not_found'):
            resultados['no_encontrados'].append(prod)
        else:
            resultados['errores'].append({'producto': prod, 'error': res.get('error','')})
        _t.sleep(0.5)  # throttle Shopify rate limit (~2 req/s)
    return jsonify({
        'ok': True,
        'procesados': len(rows),
        'sincronizados': len(resultados['ok']),
        'no_encontrados': len(resultados['no_encontrados']),
        'errores': len(resultados['errores']),
        'detalle': resultados,
    })


@bp.route('/api/formulas/sync-shopify-all', methods=['POST'])
def sync_shopify_all():
    """Sincroniza TODOS los productos pendientes en background. No bloquea.

    Util para forzar refresh masivo (ej. tras agregar productos nuevos a
    Shopify). Tambien se llama automaticamente al cargar el checklist.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db()
    token, shop = _shopify_creds(conn)
    if not token or not shop:
        return jsonify({'error': 'Shopify no configurado'}), 400
    # Contar pendientes
    pendientes = conn.execute("""
        SELECT COUNT(*) FROM formula_headers
        WHERE COALESCE(shopify_synced_at,'') = ''
           OR shopify_synced_at < datetime('now', '-5 hours', '-24 hours')
    """).fetchone()[0]
    _sync_shopify_pendientes_background(max_edad_horas=24, max_productos=100)
    return jsonify({
        'ok': True,
        'pendientes': pendientes,
        'mensaje': f'Sincronizando {pendientes} productos en background. Refresca en ~30s.'
    })

@bp.route('/api/movimientos', methods=['GET', 'POST'])
def handle_movimientos():
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        u, err, code = _require_planta_write()
        if err:
            return err, code
        data = request.get_json(silent=True) or {}
        # Sebastián 1-may-2026 audit: validar inputs antes de INSERT.
        # Antes aceptaba data sin checks · cantidad negativa, sin material_id, etc.
        material_id = (data.get('material_id') or '').strip()
        material_nombre = (data.get('material_nombre') or '').strip()
        tipo = (data.get('tipo') or '').strip()
        if not material_id:
            return jsonify({'error': 'material_id requerido'}), 400
        if not material_nombre:
            return jsonify({'error': 'material_nombre requerido'}), 400
        if tipo not in ('Entrada', 'Salida', 'Ajuste'):
            return jsonify({'error': "tipo debe ser 'Entrada', 'Salida' o 'Ajuste'"}), 400
        try:
            cantidad = float(data.get('cantidad') or 0)
        except (ValueError, TypeError):
            return jsonify({'error': 'cantidad debe ser numérico'}), 400
        if cantidad <= 0:
            return jsonify({'error': 'cantidad debe ser > 0'}), 400
        if cantidad > 1_000_000_000:  # 1 ton en gramos
            return jsonify({'error': 'cantidad excede límite razonable'}), 400
        # Pre-check de stock para Salida: rechazar si va a dejar saldo negativo.
        # Sin este check antes podía dejar stock negativo silenciosamente.
        if tipo == 'Salida':
            saldo = c.execute("""
                SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                                          WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                                          ELSE 0 END), 0)
                FROM movimientos WHERE material_id = ?
            """, (material_id,)).fetchone()[0]
            if saldo < cantidad:
                return jsonify({
                    'error': 'stock insuficiente',
                    'saldo_actual': saldo,
                    'cantidad_pedida': cantidad,
                    'sugerencia': 'Verificar stock disponible en /api/stock o registrar Entrada primero'
                }), 422
        # Sprint Movimientos PRO · 20-may-2026: lote requerido para Entrada
        # (sin lote → kardex roto, no FEFO posible). Para Salida/Ajuste sigue
        # opcional · pueden ser de conteo cíclico o ajuste sin lote específico.
        lote_in = (data.get('lote') or '').strip()
        if tipo == 'Entrada' and not lote_in:
            return jsonify({
                'error': 'lote requerido para movimientos tipo Entrada · sin lote el kardex queda roto y no se puede aplicar FEFO',
                'lote_obligatorio': True,
            }), 400
        # P2 (12-jun · hallazgo Fable): whitelist + normaliza estado_lote. Antes
        # entraba crudo -> un 'cuarentena' minuscula o un estado inventado evadia los
        # filtros NOT IN (mayusculas) del FEFO/descuento (fail-open a "usable").
        _ESTADOS_OK = ('VIGENTE', 'CUARENTENA', 'CUARENTENA_EXTENDIDA', 'RECHAZADO', 'VENCIDO', 'AGOTADO', 'BLOQUEADO')
        estado_lote_in = (data.get('estado_lote') or 'VIGENTE').strip().upper()
        if estado_lote_in not in _ESTADOS_OK:
            return jsonify({'error': 'estado_lote invalido: ' + estado_lote_in + ' · validos: ' + ', '.join(_ESTADOS_OK)}), 400
        c.execute("""INSERT INTO movimientos
                     (material_id, material_nombre, cantidad, tipo, fecha, observaciones,
                      lote, fecha_vencimiento, estanteria, posicion, proveedor, estado_lote, operador)
                     VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (material_id, material_nombre, cantidad,
                   tipo, datetime.now().isoformat(), data.get('observaciones',''),
                   lote_in, data.get('fecha_vencimiento',''),
                   data.get('estanteria',''), data.get('posicion',''),
                   data.get('proveedor',''), estado_lote_in,
                   data.get('operador','') or u))
        mov_id = c.lastrowid
        # Sprint Movimientos PRO: audit_log (faltaba en este endpoint)
        try:
            audit_log(
                c, usuario=u, accion='REGISTRAR_MOVIMIENTO_MANUAL',
                tabla='movimientos', registro_id=mov_id,
                despues={'material_id': material_id, 'tipo': tipo,
                         'cantidad': cantidad, 'lote': lote_in,
                         'observaciones': (data.get('observaciones') or '')[:200]},
                detalle=f"Movimiento manual {tipo} {cantidad}g de {material_id} · lote {lote_in or '(sin lote)'}",
            )
        except Exception as _ae:
            __import__('logging').getLogger('inventario').warning(
                'audit_log REGISTRAR_MOVIMIENTO_MANUAL fallo: %s', _ae)
        conn.commit()
        return jsonify({
            'message': 'Movimiento registrado exitosamente',
            'mov_id': mov_id, 'lote': lote_in,
        }), 201
    c.execute('SELECT id, material_id, material_nombre, cantidad, tipo, fecha, observaciones, operador, lote, proveedor, fecha_vencimiento, numero_factura FROM movimientos ORDER BY fecha DESC LIMIT 500')
    movimientos = [{'id': r[0], 'material_id': r[1] or '', 'material_nombre': r[2], 'cantidad': r[3], 'tipo': r[4], 'fecha': r[5], 'observaciones': r[6], 'operador': r[7] or '', 'lote': r[8] or '', 'proveedor': r[9] or '', 'fecha_vencimiento': r[10] or '', 'numero_factura': r[11] or ''} for r in c.fetchall()]
    return jsonify({'movimientos': movimientos})

@bp.route('/api/movimientos/<int:mov_id>', methods=['DELETE'])
def eliminar_movimiento(mov_id):
    from flask import session as flask_session
    from config import ADMIN_USERS
    usuario = flask_session.get('compras_user', '')
    if usuario not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado — solo administradores pueden eliminar movimientos'}), 403
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT id, material_id, material_nombre, lote, cantidad, tipo FROM movimientos WHERE id=?', (mov_id,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'Movimiento no encontrado'}), 404
    # audit_log obligatorio · borrar una fila del kardex es destructivo
    c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                 VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
              (usuario, 'ELIMINAR_MOVIMIENTO', 'movimientos', str(mov_id),
               f'Eliminado mov #{mov_id} ({row[5]} {row[4]} de {row[1]}) lote {row[3]}',
               request.remote_addr))
    c.execute('DELETE FROM movimientos WHERE id=?', (mov_id,))
    conn.commit()
    return jsonify({'message': f'Movimiento {mov_id} eliminado. Lote: {row[3]}, MP: {row[2]}', 'id': mov_id}), 200

@bp.route('/api/produccion', methods=['GET', 'POST'])
def handle_produccion():
    """Wrapper defensivo · captura cualquier excepción NO esperada y la
    devuelve como JSON con detalle · evita el HTML genérico 500.

    Sebastián 20-may-2026 · "no me dejo registrar produccion dice error
    interno del servidor" · si algo se nos escapa, AL MENOS sabemos qué.
    """
    try:
        return _handle_produccion_inner()
    except Exception as _outer_e:
        try:
            from flask import g as _g_outer
            _conn = getattr(_g_outer, '_conn', None) or get_db()
            try: _conn.rollback()
            except Exception: pass
        except Exception:
            pass
        import traceback as _tb
        stack = _tb.format_exc()
        __import__('logging').getLogger('inventario').exception(
            'handle_produccion · EXCEPCIÓN NO ESPERADA: %s', _outer_e)
        # Devolver MÁS info para diagnóstico inmediato cuando Sebastián
        # reporta fallas (no requiere acceso a Render logs).
        detalle = str(_outer_e) or '(excepción sin mensaje)'
        # Última línea del stack (la que tiene file:line:func)
        ultima = ''
        try:
            lines = [l for l in stack.split('\n') if l.strip()]
            for ln in reversed(lines):
                if ln.strip().startswith('File '):
                    ultima = ln.strip()[:200]
                    break
        except Exception:
            pass
        return jsonify({
            'error': 'Falla interna · revisar logs',
            'detalle': detalle[:500],
            'tipo': type(_outer_e).__name__,
            'origen': ultima,
            'rollback': 'intentado',
        }), 500


def _handle_produccion_inner():
    """Registra una producción con descuento atómico de MPs (FEFO).

    Flujo robusto (todo o nada):
      1. VALIDAR input (cantidad > 0, producto no vacío, fórmula existe).
      2. PRE-CHECK stock: para cada MP, consultar lotes disponibles (excluyendo
         CUARENTENA/RECHAZADO) y verificar que la suma cubre el requerimiento.
         Si falta stock para CUALQUIER MP → 422 sin escribir nada.
      3. EJECUTAR transacción: INSERT producciones + INSERT movimientos por lote.
         Si algo falla → ROLLBACK explícito + log + 500 con detalle.

    Esto reemplaza el flujo anterior que (a) creaba "salida sin lote" cuando no
    había stock — generando stock negativo silencioso, y (b) no validaba que la
    fórmula existiera — registraba producciones sin descuentos.

    Excepción: si una MP está marcada como ilimitada (agua, etc.) en
    programacion._MP_UNLIMITED, el pre-check la salta. Se sigue registrando el
    movimiento de salida (para trazabilidad de pesaje) pero sin requerir stock.
    """
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        u, err, code = _require_planta_write()
        if err:
            return err, code
        # SEC-FIX · 21-may-2026 · BEGIN IMMEDIATE atómico (race FEFO multi-worker)
        # Antes: pre-check + INSERT con 3 workers Gunicorn · dos producciones
        # concurrentes del mismo MP pasaban pre-check ambos → stock negativo
        # silencioso. Ahora: BEGIN IMMEDIATE bloquea otras escrituras hasta
        # commit · sin riesgo de race (cost: serialización por DB lock).
        try:
            conn.execute('BEGIN IMMEDIATE')
        except Exception:
            pass  # PG no soporta · usa SERIALIZABLE implícito

        data = request.json or {}
        producto = (data.get('producto') or '').strip()
        presentacion = data.get('presentacion','')
        cantidad_kg = float(data.get('cantidad_kg', data.get('cantidad', 0)) or 0)
        cantidad_g = cantidad_kg * 1000
        operador = (data.get('operador') or '').strip()
        observaciones_in = data.get('observaciones', '')
        # N° de Lote bulk (MyBatch parity): si el operario lo escribe, se usa ese;
        # si lo deja vacío, EOS auto-genera PROD-NNNNN. Sanitizado básico.
        lote_in = (data.get('lote') or '').strip()[:40]

        # ─── Validación 1: input ─────────────────────────────────────────────
        if not producto:
            return jsonify({'error': 'Producto vacío'}), 400
        if cantidad_kg <= 0:
            return jsonify({'error': 'cantidad_kg debe ser > 0'}), 400
        if not operador:
            return jsonify({'error': 'Falta operador'}), 400

        # ─── Validación 2: fórmula existe ───────────────────────────────────
        c.execute(
            # FIX 13-jun (audit fórmulas · GMP): NO fabricar desde una fórmula
            # DESCONTINUADA. Una fórmula con activo=0 (descontinuada · mig 229/230/231)
            # conserva sus formula_items; sin este filtro, producir con su nombre exacto
            # descontaba la fórmula vieja/incompleta (caso 'Blush Balm' minúscula 67% vs
            # 'BLUSH BALM' completa). Excluye solo headers EXPLÍCITAMENTE activo=0 (ningún
            # nombre exacto tiene header activo+inactivo a la vez · verificado).
            'SELECT material_id, material_nombre, porcentaje FROM formula_items '
            'WHERE producto_nombre=? AND producto_nombre NOT IN '
            '(SELECT producto_nombre FROM formula_headers WHERE COALESCE(activo,1)=0)',
            (producto,)
        )
        formula_items = c.fetchall()
        if not formula_items:
            return jsonify({
                'error': f"Producto sin fórmula registrada: '{producto}'",
                'detalle': 'Crea la fórmula en /tecnica antes de producir'
            }), 400

        # ─── Validación 3: idempotencia (anti-doble-click / replay) ─────────
        # Si en los últimos 90 segundos ya se registró la MISMA producción
        # (mismo producto + cantidad + operador), devolvemos la existente
        # en lugar de crear una nueva. Cubre el caso clásico de doble-click,
        # cliente con red lenta que reintenta, o un POST replay.
        try:
            c.execute("""SELECT id, fecha, lote
                         FROM producciones
                         WHERE producto=? AND cantidad=? AND operador=?
                           AND datetime(fecha) >= datetime('now', '-5 hours', '-90 seconds')
                         ORDER BY id DESC LIMIT 1""",
                      (producto, cantidad_kg, operador))
            dup = c.fetchone()
            if dup:
                return jsonify({
                    'message': 'Producción ya registrada hace <90s — duplicado evitado',
                    'lote': dup[2] or f'PROD-{dup[0]:05d}',
                    'duplicado': True,
                    'id_existente': dup[0],
                }), 200
        except sqlite3.OperationalError:
            pass  # esquema antiguo sin alguna columna — continuar normal

        # ─── MPs ilimitadas (no requieren validación de stock) ──────────────
        # Carga la lista desde programacion (única fuente de verdad)
        try:
            from .programacion import _MP_UNLIMITED, _norm_mp_name
            unlimited_set = set(_MP_UNLIMITED)
            def _is_unlimited(nombre):
                return _norm_mp_name(nombre or '').upper() in {x.upper() for x in unlimited_set}
        except Exception:
            unlimited_set = set()
            def _is_unlimited(nombre):  # noqa
                return False

        # Audit 4-jun · MISMA resolución código fórmula→bodega que simular y el
        # descuento canónico (M1 · resolver único). Antes este endpoint validaba
        # contra el código CRUDO de fórmula → veía 0g cuando el stock está bajo un
        # código duplicado/inactivo (Terpenos→MP00181, Pantenol→código inactivo) y
        # abortaba aunque "Verificar Stock" mostrara el inventario OK.
        try:
            from .programacion import _resolver_material_bodega as _resolver_prod
        except Exception:
            _resolver_prod = None

        # MP infinita / fabricada en casa (AGUA del lab, mig 218 controla_stock=0):
        # no exige ni descuenta stock. Complementa la lista _MP_UNLIMITED por nombre.
        def _no_controla(cod):
            if not cod:
                return False
            try:
                r = c.execute("SELECT COALESCE(controla_stock,1) FROM maestro_mps WHERE codigo_mp=?", (cod,)).fetchone()
                return r is not None and int(r[0] or 0) == 0
            except Exception:
                return False

        # ─── PRE-CHECK: stock suficiente para TODAS las MPs ─────────────────
        # Construimos el plan completo SIN escribir nada. Si falta stock para
        # alguna MP, abortamos antes del primer INSERT.
        plan_descuentos = []  # cada entry: {mat_id, mat_nombre, g_total, lotes_a_usar:[(lote, vence, g)]}
        faltantes = []        # MPs que no tienen stock suficiente
        for mat_id, mat_nombre, pct in formula_items:
            g_total = round((pct / 100.0) * cantidad_g, 2)
            if g_total <= 0:
                continue
            # Resolver código fórmula → bodega (mismo material en código duplicado/
            # inactivo CON stock). Todas las consultas y la Salida usan el resuelto.
            mat_id_formula = mat_id
            try:
                if _resolver_prod:
                    mat_id = _resolver_prod(c, mat_id, mat_nombre) or mat_id
            except Exception:
                mat_id = mat_id_formula
            entry = {
                'mat_id': mat_id, 'mat_nombre': mat_nombre,
                'g_total': g_total, 'lotes_a_usar': [],
                'g_sin_lote': 0.0, 'unlimited': False,
            }
            if _is_unlimited(mat_nombre) or _no_controla(mat_id) or _no_controla(mat_id_formula):
                # Aguas y similares (lista _MP_UNLIMITED o controla_stock=0) —
                # pesaje real pero sin requerir stock.
                entry['unlimited'] = True
                if _no_controla(mat_id) or _no_controla(mat_id_formula):
                    # AGUA del lab (infinita, fabricada en casa): NO mover kardex
                    # para no acumular stock negativo (-330k g). No se descuenta.
                    entry['g_sin_lote'] = 0.0
                else:
                    # _MP_UNLIMITED legacy por nombre: deja Salida de pesaje (trazabilidad).
                    entry['g_sin_lote'] = g_total
                plan_descuentos.append(entry)
                continue

            # FEFO sobre lotes disponibles (excluye CUARENTENA/RECHAZADO)
            # Sebastian 8-may-2026 zero-error: tomar fv REAL de la Entrada
            # original (no del primer mov que SQLite escoja arbitrariamente).
            # Sebastián 20-may-2026 HOTFIX: en PostgreSQL `fv_real = ''`
            # falla con error de tipo si fecha_vencimiento es columna DATE
            # (no TEXT). Cambiado a COALESCE + CAST a TEXT que funciona en
            # ambos. Antes "no me dejo registrar producción" 500 silencioso.
            c.execute("""SELECT lote,
                                MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) AS fv_real,
                                SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock
                         FROM movimientos
                         WHERE material_id=? AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                           AND (estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO'))
                         GROUP BY lote HAVING stock > 0.01
                           AND (fv_real IS NULL OR TRIM(CAST(fv_real AS TEXT))='' OR date(fv_real) >= date('now', '-5 hours'))
                         ORDER BY COALESCE(NULLIF(CAST(MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) AS TEXT), ''), '9999-12-31') ASC""", (mat_id,))
            lotes_fefo = c.fetchall()
            stock_total_disp = sum(float(l[2] or 0) for l in lotes_fefo)
            if stock_total_disp + 0.01 < g_total:  # tolerancia 0.01g por floats
                # 2-jun-2026 · TRANSPARENCIA: ¿cuánto stock de este código está
                # RETENIDO en estados no-producibles (VENCIDO/AGOTADO/CUARENTENA)?
                # Caso CONTORNO: 600g del lote YT20251203 marcado no-usable → por
                # eso producción ve 17.5g aunque Bodega muestre 617g.
                retenido = {}
                ret_total = 0.0
                try:
                    for er in c.execute(
                        """SELECT COALESCE(estado_lote,'(sin estado)') AS est, lote,
                                  SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stk
                           FROM movimientos
                           WHERE material_id=? AND estado_lote IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO')
                           -- AS explícito en stk: el reescritor de HAVING-alias del adaptador
                           -- solo expande alias con AS; sin él, `HAVING stk` daba
                           -- "column stk does not exist" en PG → retenido vacío. Suite PG 8-jun.
                           GROUP BY estado_lote, lote HAVING stk > 0""", (mat_id,)).fetchall():
                        g = float(er[2] or 0)
                        retenido.setdefault(er[0], 0.0)
                        retenido[er[0]] += g
                        ret_total += g
                    retenido = {k: round(v, 2) for k, v in retenido.items()}
                except Exception:
                    retenido = {}
                faltantes.append({
                    'material': mat_nombre,
                    'material_id': mat_id,
                    'requerido_g': g_total,
                    'disponible_g': round(stock_total_disp, 2),
                    'falta_g': round(g_total - stock_total_disp, 2),
                    'retenido_g': round(ret_total, 2),
                    'retenido_por_estado': retenido,
                })
                continue

            # Plan FEFO: cuánto sacar de cada lote
            g_restante = g_total
            for lrow in lotes_fefo:
                if g_restante <= 0:
                    break
                lote_n, lote_v, lote_s = lrow
                g_lote = round(min(g_restante, float(lote_s)), 2)
                entry['lotes_a_usar'].append({
                    'lote': lote_n,
                    'vence': str(lote_v)[:10] if lote_v else '',
                    'g': g_lote,
                })
                g_restante = round(g_restante - g_lote, 2)
            plan_descuentos.append(entry)

        if faltantes:
            # Sebastián 21-may-2026: AUTO-DETECTAR codigo_mp huérfanos en el
            # mismo response · si la fórmula apunta a un código sin stock
            # pero existe MP con nombre similar QUE SÍ tiene stock, ofrecer
            # auto-repair sin tener que apretar otro botón.
            auto_repair_candidatos = []
            for f in faltantes:
                mid = f.get('material_id', '')
                mnom = (f.get('material', '') or '').strip()
                if not mid or not mnom:
                    continue
                nom_norm = mnom.lower()[:30]
                if len(nom_norm) < 4:
                    continue
                try:
                    cand = c.execute(
                        """SELECT mp.codigo_mp, mp.nombre_comercial,
                                  COALESCE(SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END), 0) as stock_act
                           FROM maestro_mps mp
                           LEFT JOIN movimientos m ON m.material_id = mp.codigo_mp
                             AND (m.estado_lote IS NULL OR m.estado_lote NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO'))
                           WHERE COALESCE(mp.activo,1)=1
                             AND mp.codigo_mp != ?
                             AND (LOWER(COALESCE(mp.nombre_comercial,'')) LIKE ?
                                  OR LOWER(COALESCE(mp.nombre_inci,'')) LIKE ?)
                           GROUP BY mp.codigo_mp, mp.nombre_comercial
                           HAVING stock_act > 0
                           ORDER BY stock_act DESC LIMIT 1""",
                        (mid, f'%{nom_norm}%', f'%{nom_norm}%'),
                    ).fetchone()
                except Exception:
                    cand = None
                if cand:
                    auto_repair_candidatos.append({
                        'huerfano': {'codigo': mid, 'nombre': mnom},
                        'reemplazo': {'codigo': cand[0], 'nombre': cand[1],
                                      'stock_g': float(cand[2] or 0)},
                    })
            return jsonify({
                'error': 'Stock insuficiente para producir',
                'producto': producto,
                'cantidad_kg': cantidad_kg,
                'faltantes': faltantes,
                'auto_repair_candidatos': auto_repair_candidatos,
                'auto_repair_disponible': len(auto_repair_candidatos) > 0,
                'mensaje': (
                    f"No se puede producir {cantidad_kg}kg de {producto}: "
                    f"{len(faltantes)} MP(s) sin stock suficiente. "
                    + (f"Detectados {len(auto_repair_candidatos)} codigo_mp huérfanos · "
                       "usar botón 🔧 Auto-reparar." if auto_repair_candidatos else
                       "Verifica entradas en Bodega MP o crea OC en /compras.")
                ),
            }), 422  # Unprocessable Entity

        # ─── ESCRITURA ATÓMICA ──────────────────────────────────────────────
        # SQLite con isolation_level='DEFERRED' (default): primer DML inicia
        # transacción implícita, conn.commit() la cierra, conn.rollback() la
        # descarta. Si una excepción ocurre antes del commit, los inserts
        # quedan en transacción y se rollbackean al cerrar la conexión.
        # Aún así, hacemos rollback EXPLÍCITO para no depender de timing.
        fecha = datetime.now().isoformat()
        prod_id = None
        lote_ref = None
        descuentos = []
        try:
            c.execute(
                'INSERT INTO producciones (producto, cantidad, fecha, estado, observaciones, operador, presentacion) VALUES (?,?,?,?,?,?,?)',
                (producto, cantidad_kg, fecha, 'Completado', observaciones_in, operador, presentacion)
            )
            prod_id = c.lastrowid
            lote_ref = lote_in if lote_in else f'PROD-{prod_id:05d}'
            try:
                c.execute("UPDATE producciones SET lote=? WHERE id=?", (lote_ref, prod_id))
            except sqlite3.OperationalError as _e:
                if 'no such column' not in str(_e).lower():
                    raise

            # Sebastián 8-may-2026 zero-error: snapshot inmutable de la fórmula
            # al momento exacto. Previene drift retroactivo: si la fórmula se
            # modifica después, audit compara contra esta versión.
            try:
                import json as _json_snap
                items_snap = c.execute(
                    "SELECT material_id, material_nombre, porcentaje "
                    "FROM formula_items WHERE producto_nombre=? AND producto_nombre NOT IN "
                    "(SELECT producto_nombre FROM formula_headers WHERE COALESCE(activo,1)=0)",
                    (producto,)
                ).fetchall()
                snap = [{'material_id': r[0], 'material_nombre': r[1],
                         'porcentaje': r[2]} for r in items_snap]
                c.execute(
                    "UPDATE producciones SET formula_snapshot_json=? WHERE id=?",
                    (_json_snap.dumps(snap), prod_id)
                )
            except sqlite3.OperationalError as _e:
                # Columna formula_snapshot_json puede no existir en schemas viejos
                if 'no such column' not in str(_e).lower():
                    raise

            for plan in plan_descuentos:
                lotes_log = []
                for uso in plan['lotes_a_usar']:
                    # FIX 11-jun (drift PG) · NO insertar un movimiento de 0/negativo:
                    # el trigger fn_trg_mov_cantidad_positiva lo rechaza y aborta TODA
                    # la producción (HTTP 500). Si la distribución FEFO asignó 0 g a un
                    # lote (redondeo / lote sin saldo), se salta · otros lotes cubren el total.
                    if float(uso.get('g') or 0) <= 0:
                        continue
                    c.execute(
                        "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, observaciones, lote, operador) VALUES (?,?,?,?,?,?,?,?)",
                        (plan['mat_id'], plan['mat_nombre'], uso['g'], 'Salida', fecha,
                         f"FEFO:{lote_ref}:{producto} x {cantidad_kg}kg", uso['lote'], operador)
                    )
                    lotes_log.append({'lote': uso['lote'], 'vence': uso['vence'], 'cantidad_g': uso['g']})
                # MPs ilimitadas: registrar movimiento de pesaje sin lote
                if plan.get('unlimited') and plan['g_sin_lote'] > 0:
                    c.execute(
                        "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, observaciones, operador) VALUES (?,?,?,?,?,?,?)",
                        (plan['mat_id'], plan['mat_nombre'], plan['g_sin_lote'], 'Salida', fecha,
                         f"UNLIMITED:{lote_ref}:{producto} x {cantidad_kg}kg (MP sin requerir stock)", operador)
                    )
                    lotes_log.append({'lote': 'unlimited', 'vence': '', 'cantidad_g': plan['g_sin_lote']})

                descuentos.append({
                    'material': plan['mat_nombre'],
                    'material_id': plan['mat_id'],
                    'cantidad_g': plan['g_total'],
                    'unlimited': plan.get('unlimited', False),
                    'lotes_fefo': lotes_log,
                })

            # P0-6 23-may-PM · auditoría agente Stock · _handle_produccion_inner
            # descontaba MP vía INSERT en movimientos pero NO escribía
            # audit_log · viola CLAUDE.md 'audit_log es mandatorio en cualquier
            # operación que mutue inventario'. Imposible reconstruir quién
            # descontó qué lote para reclamo INVIMA.
            try:
                from database import audit_log as _al
                _al(c, usuario=operador or 'sistema',
                    accion='PRODUCCION_DESCONTAR_MP',
                    tabla='movimientos', registro_id=str(lote_ref),
                    despues={'producto': producto,
                             'cantidad_kg': cantidad_kg,
                             'descuentos': descuentos,
                             'fecha': fecha})
            except Exception:
                pass

            # Reemplazo MyBatch · 5-jun-2026 · LEGAJO AUTOMÁTICO (como MyBatch).
            # Al registrar producción se crea el EBR (batch record) SIEMPRE que el
            # producto tenga un MBR APROBADO. crear_ebr_desde_mbr es auto-gateado:
            # si NO hay MBR aprobado devuelve NO_MBR_APROBADO y no hace nada (no
            # bloquea). NO depende de EBR_MODE: el legajo nace solo sin forzar la
            # e-firma de pesajes (eso lo controla EBR_MODE aparte). Cero riesgo:
            # productos sin MBR aprobado se comportan idéntico a antes.
            ebr_auto = None
            try:
                from blueprints.brd import crear_ebr_desde_mbr
                _area_cod = (data.get('area_codigo') or '').strip()
                _r = crear_ebr_desde_mbr(
                    c, producto_nombre=producto, lote=lote_ref,
                    produccion_id=None, cantidad_objetivo_g=cantidad_g,
                    usuario=operador, area_codigo=_area_cod)
                if _r.get('ok'):
                    ebr_auto = _r
                    try:
                        from database import audit_log as _al2
                        _al2(c, usuario=operador or 'sistema', accion='CREAR_EBR_AUTO',
                             tabla='ebr_ejecuciones', registro_id=str(_r.get('id')),
                             despues={'producto': producto, 'lote': lote_ref,
                                      'numero_op': _r.get('numero_op')})
                    except Exception:
                        pass
            except Exception as _eebr:
                __import__('logging').getLogger('inventario').warning(
                    'crear EBR auto en registro fallo (no bloquea producción): %s', _eebr)

            conn.commit()
        except Exception as _e:
            conn.rollback()
            __import__('logging').getLogger('inventario').error(
                "Producción FALLÓ tras pre-check OK (rollback): producto=%s kg=%s err=%s",
                producto, cantidad_kg, _e, exc_info=True
            )
            return jsonify({
                'error': 'Falla transaccional al registrar producción',
                'detalle': str(_e),
                'rollback': 'aplicado — no se descontó nada y no quedó producción registrada',
            }), 500

        # Sprint Fabricación PRO 20-may-2026: persistir costo estimado
        try:
            costo_total_cop = 0.0
            for plan in plan_descuentos:
                if plan.get('unlimited'):
                    continue
                # Buscar precio_referencia (COP/kg) en maestro_mps
                pr_row = c.execute(
                    "SELECT COALESCE(precio_referencia, 0) FROM maestro_mps "
                    "WHERE codigo_mp=? LIMIT 1", (plan['mat_id'],),
                ).fetchone()
                if pr_row and pr_row[0]:
                    # precio es por kg · cantidad es en g
                    costo_total_cop += float(pr_row[0]) * (plan['g_total'] / 1000.0)
            if costo_total_cop > 0:
                try:
                    c.execute(
                        "UPDATE producciones SET costo_estimado_cop=? WHERE id=?",
                        (round(costo_total_cop, 2), prod_id),
                    )
                    conn.commit()
                except Exception as _e:
                    # PostgreSQL: psycopg2.UndefinedColumn · SQLite: OperationalError
                    # Si la columna no existe (mig 148 no aplicada), no romper
                    # el flujo y rollback para evitar tx zombie en PG.
                    err_msg = str(_e).lower()
                    if 'no such column' in err_msg or 'undefined column' in err_msg or 'does not exist' in err_msg:
                        try: conn.rollback()
                        except Exception: pass
                    else:
                        try: conn.rollback()
                        except Exception: pass
                        __import__('logging').getLogger('inventario').warning(
                            'costo_estimado_cop UPDATE fallo no-schema: %s', _e)
        except Exception as _e_outer:
            # Defensivo: cualquier otro error en cálculo de costo NO debe
            # romper la respuesta del POST. Loguear y seguir.
            try: conn.rollback()
            except Exception: pass
            __import__('logging').getLogger('inventario').warning(
                'costo_estimado_cop calc fallo: %s', _e_outer)

        # 15-jun · CÁLCULO PERFECTO · espejo automático Fabricación → calendario.
        # Cada producción registrada se refleja como lote COMPLETADO retroactivo en
        # produccion_programada para que el ANCLA del cálculo (ultima_prod) la cuente
        # y aparezca en el calendario, SIN backfill manual. Best-effort: la producción
        # YA está commiteada · si el espejo falla, jamás rompe el registro (el cron
        # job_sync_fabricacion_calendario lo reconcilia después). M22: audit + commit.
        try:
            from blueprints.plan import _mirror_produccion_a_calendario as _mirror_fab
            if _mirror_fab(conn, prod_id, producto, cantidad_kg, fecha, lote_ref,
                           usuario=(operador or 'fabricacion')):
                conn.commit()
        except Exception as _emir:
            try:
                conn.rollback()
            except Exception:
                pass
            __import__('logging').getLogger('inventario').warning(
                'espejo Fabricación→calendario fallo (no bloquea · cron reconcilia): %s', _emir)

        # Stock PT se crea via Acondicionamiento → Liberacion (flujo BPM correcto)
        msg = f'Produccion registrada: {producto} x {cantidad_kg}kg (FEFO)'
        if descuentos:
            msg += f'. {len(descuentos)} MPs descontadas.'
        if ebr_auto and ebr_auto.get('numero_op'):
            msg += f' · Legajo {ebr_auto["numero_op"]} creado automáticamente.'
        return jsonify({'message': msg, 'descuentos': descuentos, 'lote': lote_ref,
                        'ebr': ebr_auto}), 201
    # Sprint Fabricación PRO 20-may-2026: paginación + búsqueda + filtros
    # server-side. Antes solo LIMIT 50 sin offset ni q · imposible navegar.
    try:
        limit = max(1, min(int(request.args.get('limit', 50)), 500))
    except (ValueError, TypeError):
        limit = 50
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    q = (request.args.get('q') or '').strip()
    desde = (request.args.get('desde') or '').strip()
    hasta = (request.args.get('hasta') or '').strip()
    where_parts = ['1=1']
    params = []
    if q:
        # Escapar % y _ para LIKE
        qesc = q.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
        where_parts.append("(LOWER(producto) LIKE LOWER(?) ESCAPE '\\' OR LOWER(COALESCE(lote,'')) LIKE LOWER(?) ESCAPE '\\' OR LOWER(COALESCE(operador,'')) LIKE LOWER(?) ESCAPE '\\')")
        params += [f'%{qesc}%', f'%{qesc}%', f'%{qesc}%']
    if desde:
        where_parts.append("fecha >= ?"); params.append(desde)
    if hasta:
        where_parts.append("fecha <= ?"); params.append(hasta + ' 23:59:59')
    where_sql = ' AND '.join(where_parts)
    # Total count para paginación
    try:
        total_row = c.execute(f"SELECT COUNT(*) FROM producciones WHERE {where_sql}", params).fetchone()
        total = int(total_row[0]) if total_row else 0
    except Exception:
        total = 0
    # Query principal (COALESCE columna costo_estimado_cop puede no existir aún en BDs sin mig 148)
    try:
        rows = c.execute(
            f"""SELECT id, producto, cantidad, fecha, estado, operador,
                       COALESCE(presentacion,'') AS pres,
                       COALESCE(lote,'') AS lote,
                       COALESCE(observaciones,'') AS obs,
                       COALESCE(costo_estimado_cop, 0) AS costo
                FROM producciones
                WHERE {where_sql}
                ORDER BY fecha DESC, id DESC
                LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
    except Exception as _e_main:
        # PG/SQLite agnostic · si columna costo_estimado_cop no existe,
        # rollback la tx zombie y reintentar sin esa columna.
        try: conn.rollback()
        except Exception: pass
        rows = c.execute(
            f"""SELECT id, producto, cantidad, fecha, estado, operador,
                       COALESCE(presentacion,'') AS pres,
                       COALESCE(lote,'') AS lote,
                       COALESCE(observaciones,'') AS obs,
                       0 AS costo
                FROM producciones
                WHERE {where_sql}
                ORDER BY fecha DESC, id DESC
                LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
    prod = [{
        'id': r[0],
        'lote': r[7] or f'PROD-{r[0]:05d}',
        'producto': r[1], 'cantidad': r[2], 'fecha': r[3],
        'estado': r[4], 'operador': r[5] or '',
        'presentacion': r[6] or '',
        'observaciones': r[8] or '',
        'costo_estimado_cop': float(r[9] or 0),
    } for r in rows]
    return jsonify({
        'producciones': prod,
        'total': total, 'limit': limit, 'offset': offset,
        'q': q, 'desde': desde, 'hasta': hasta,
    })


@bp.route('/api/produccion/pendientes-hoy', methods=['GET'])
def produccion_pendientes_hoy():
    """Sprint Fabricación PRO · 20-may-2026 · banner "Pendientes hoy".

    Lista producciones programadas para HOY que aún no se iniciaron ni
    completaron · sirve para que Luis Enrique/Mayerlin sepan qué falta
    al abrir la pestaña Fabricación.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    try:
        rows = c.execute(
            """SELECT id, producto, COALESCE(cantidad_kg, 0)
               FROM produccion_programada
               WHERE date(fecha_programada) = date('now','-5 hours')
                 AND COALESCE(estado, 'programado') NOT IN ('cancelado', 'completado')
                 AND inicio_real_at IS NULL
               ORDER BY id LIMIT 20""",
        ).fetchall()
    except Exception:
        rows = []
    items = [{'id': r[0], 'producto': r[1], 'kg': float(r[2] or 0)} for r in rows]
    return jsonify({'items': items, 'count': len(items)})


@bp.route('/api/produccion/<int:pid>/detalle', methods=['GET'])
def produccion_detalle(pid):
    """Sprint Fabricación PRO · detalle completo de una producción para
    el modal "Ver detalle" del historial.

    Devuelve: header + MPs descontadas con lotes FEFO usados + costo +
    snapshot fórmula al momento.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    try:
        h = c.execute(
            """SELECT id, producto, cantidad, fecha, estado, operador,
                      COALESCE(presentacion,''), COALESCE(lote,''),
                      COALESCE(observaciones,''),
                      COALESCE(formula_snapshot_json,''),
                      COALESCE(costo_estimado_cop, 0)
               FROM producciones WHERE id=?""",
            (pid,),
        ).fetchone()
    except Exception:
        # PG/SQLite agnostic: fallback sin costo_estimado_cop si la
        # mig 148 no aplicó. Rollback tx zombie en PG.
        try: conn.rollback()
        except Exception: pass
        h = c.execute(
            """SELECT id, producto, cantidad, fecha, estado, operador,
                      COALESCE(presentacion,''), COALESCE(lote,''),
                      COALESCE(observaciones,''),
                      '' AS snap, 0 AS costo
               FROM producciones WHERE id=?""",
            (pid,),
        ).fetchone()
    if not h:
        return jsonify({'error': 'producción no existe'}), 404
    lote_ref = h[7] or f'PROD-{h[0]:05d}'
    # Movimientos de salida vinculados a esta producción (por observaciones)
    movs = c.execute(
        """SELECT material_id, material_nombre, cantidad, lote,
                  COALESCE(observaciones,'')
           FROM movimientos
           WHERE tipo='Salida'
             AND (observaciones LIKE ? OR observaciones LIKE ?)
           ORDER BY id""",
        (f'%FEFO:{lote_ref}:%', f'%UNLIMITED:{lote_ref}:%'),
    ).fetchall()
    descuentos = [{
        'material_id': r[0], 'material_nombre': r[1],
        'cantidad_g': float(r[2] or 0),
        'lote': r[3] or 'unlimited',
        'observaciones': r[4],
    } for r in movs]
    snap = []
    if h[9]:
        try:
            import json as _json
            snap = _json.loads(h[9])
        except Exception:
            snap = []
    return jsonify({
        'id': h[0], 'producto': h[1], 'cantidad_kg': h[2],
        'fecha': h[3], 'estado': h[4], 'operador': h[5] or '',
        'presentacion': h[6], 'lote': lote_ref,
        'observaciones': h[8],
        'costo_estimado_cop': float(h[10] or 0),
        'descuentos': descuentos,
        'formula_snapshot': snap,
    })


@bp.route('/api/produccion/auditar-formulas-huerfanas', methods=['GET'])
def produccion_auditar_formulas_huerfanas():
    """Sprint Fabricación PRO · 21-may-2026.

    Recorre TODAS las fórmulas y detecta material_id huérfanos
    (sin stock FEFO disponible) que tienen UN candidato claro por
    nombre similar (otro codigo_mp con stock).

    Útil para reparar masivo post-unificación de duplicados.
    Solo lectura (no aplica nada · dry-run global).
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    rows = c.execute(
        """SELECT producto_nombre, material_id, material_nombre
           FROM formula_items
           WHERE material_id IS NOT NULL AND material_id != ''
           ORDER BY producto_nombre, material_id""",
    ).fetchall()
    formulas_afectadas = {}
    for prod, mid, mnom in rows:
        # ¿Stock FEFO usable?
        try:
            r = c.execute(
                """SELECT COALESCE(SUM(stock_t),0) FROM (
                     SELECT lote, SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_t
                     FROM movimientos
                     WHERE material_id=? AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                       AND (estado_lote IS NULL OR estado_lote NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO'))
                     GROUP BY lote HAVING stock_t > 0)""",
                (mid,),
            ).fetchone()
            stock_fefo = float(r[0] or 0) if r else 0
        except Exception:
            stock_fefo = 0
        if stock_fefo > 0:
            continue  # OK
        # Buscar candidato
        nom_norm = (mnom or '').strip().lower()
        if len(nom_norm) < 4:
            continue
        try:
            cand = c.execute(
                """SELECT mp.codigo_mp, mp.nombre_comercial,
                          COALESCE(SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END), 0) as stock_act
                   FROM maestro_mps mp
                   LEFT JOIN movimientos m ON m.material_id = mp.codigo_mp
                     AND (m.estado_lote IS NULL OR UPPER(COALESCE(m.estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))
                   WHERE COALESCE(mp.activo,1)=1
                     AND mp.codigo_mp != ?
                     AND (LOWER(COALESCE(mp.nombre_comercial,'')) LIKE ?
                          OR LOWER(COALESCE(mp.nombre_inci,'')) LIKE ?)
                   GROUP BY mp.codigo_mp, mp.nombre_comercial
                   HAVING stock_act > 0
                   ORDER BY stock_act DESC LIMIT 1""",
                (mid, f'%{nom_norm[:30]}%', f'%{nom_norm[:30]}%'),
            ).fetchone()
        except Exception:
            cand = None
        if not cand:
            continue
        formulas_afectadas.setdefault(prod, []).append({
            'huerfano': {'codigo': mid, 'nombre': mnom},
            'reemplazo': {'codigo': cand[0], 'nombre': cand[1],
                          'stock_g': float(cand[2] or 0)},
        })
    return jsonify({
        'productos_afectados_count': len(formulas_afectadas),
        'total_huerfanos': sum(len(v) for v in formulas_afectadas.values()),
        'productos': [{'producto': k, 'cambios': v} for k, v in formulas_afectadas.items()],
        'mensaje': (
            f'{len(formulas_afectadas)} fórmula(s) con codigo_mp huérfanos · '
            f'{sum(len(v) for v in formulas_afectadas.values())} cambios sugeridos · '
            'usar /api/produccion/auto-reparar-todas para aplicar masivo (admin).'
        ),
    })


@bp.route('/api/produccion/auto-reparar-todas', methods=['POST'])
def produccion_auto_reparar_todas():
    """Sprint Fabricación PRO · 21-may-2026 · admin · aplica auto-reparar
    a TODAS las fórmulas con codigo_mp huérfanos detectables.

    Body: {dry_run: bool} · default true.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if (u or '').strip().lower() not in {x.lower() for x in ADMIN_USERS}:
        return jsonify({'error': 'solo admin'}), 403
    body = request.get_json(silent=True) or {}
    dry_run = body.get('dry_run', True) is True
    # Re-usar la auditoría
    import json as _json
    from flask import url_for
    # Llamamos al endpoint internamente para no duplicar lógica
    audit_resp = produccion_auditar_formulas_huerfanas()
    if hasattr(audit_resp, 'get_json'):
        audit_data = audit_resp.get_json()
    else:
        audit_data = _json.loads(audit_resp[0].data) if isinstance(audit_resp, tuple) else {}
    productos = audit_data.get('productos', [])
    conn = get_db(); c = conn.cursor()
    aplicados = []
    if not dry_run:
        for p_info in productos:
            prod = p_info['producto']
            for ch in p_info['cambios']:
                try:
                    c.execute(
                        "UPDATE formula_items SET material_id=?, material_nombre=? "
                        "WHERE producto_nombre=? AND material_id=?",
                        (ch['reemplazo']['codigo'], ch['reemplazo']['nombre'],
                         prod, ch['huerfano']['codigo']),
                    )
                    aplicados.append({
                        'producto': prod,
                        'huerfano': ch['huerfano']['codigo'],
                        'reemplazo': ch['reemplazo']['codigo'],
                    })
                except Exception:
                    continue
        try:
            audit_log(c, usuario=u, accion='AUTO_REPARAR_FORMULAS_BULK',
                      tabla='formula_items', registro_id='_BULK_',
                      despues={'aplicados': len(aplicados)})
        except Exception:
            pass
        conn.commit()
    return jsonify({
        'dry_run': dry_run,
        'productos_con_cambios': len(productos),
        'cambios_totales': audit_data.get('total_huerfanos', 0),
        'aplicados': aplicados,
        'mensaje': (
            f'{audit_data.get("total_huerfanos",0)} cambio(s) propuestos en {len(productos)} fórmula(s)'
            if dry_run else f'{len(aplicados)} cambio(s) aplicado(s)'
        ),
    })


@bp.route('/api/produccion/auto-reparar-formula/<path:producto>', methods=['POST'])
def produccion_auto_reparar_formula(producto):
    """Sprint Fabricación PRO · 20-may-2026.

    Si una fórmula tiene material_id huérfanos (sin lotes ni stock pero
    hay un código MP similar con stock real), ofrece reemplazar el código
    huérfano por el verdadero, basado en match por nombre normalizado.

    Body: {dry_run: bool}  dry_run=true (default) → preview · false → aplica
    Solo admin.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if (u or '').strip().lower() not in {x.lower() for x in ADMIN_USERS}:
        return jsonify({'error': 'solo admin puede auto-reparar fórmulas'}), 403
    body = request.get_json(silent=True) or {}
    dry_run = body.get('dry_run', True) is True
    conn = get_db(); c = conn.cursor()
    items = c.execute(
        "SELECT material_id, material_nombre, porcentaje FROM formula_items WHERE producto_nombre=?",
        (producto,),
    ).fetchall()
    if not items:
        return jsonify({'error': f'producto "{producto}" sin fórmula'}), 404
    cambios_propuestos = []
    aplicados = []
    for mid, mnom, pct in items:
        # ¿Tiene stock FEFO?
        try:
            row = c.execute(
                """SELECT COALESCE(SUM(stock_t),0) FROM (
                     SELECT lote, SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_t
                     FROM movimientos
                     WHERE material_id=? AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                       AND (estado_lote IS NULL OR estado_lote NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO'))
                     GROUP BY lote HAVING stock_t > 0)""",
                (mid,),
            ).fetchone()
            stock_fefo = float(row[0] or 0) if row else 0
        except Exception:
            stock_fefo = 0
        if stock_fefo > 0:
            continue  # OK, no necesita reparación
        # Buscar candidatos: codigo_mp con nombre normalizado similar Y con stock
        nom_norm = (mnom or '').strip().lower()
        if len(nom_norm) < 4:
            continue
        candidatos = c.execute(
            """SELECT mp.codigo_mp, mp.nombre_comercial,
                      COALESCE(SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END), 0) as stock_act
               FROM maestro_mps mp
               LEFT JOIN movimientos m ON m.material_id = mp.codigo_mp
                 AND (m.estado_lote IS NULL OR UPPER(COALESCE(m.estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','VENCIDO','RECHAZADO','AGOTADO','BLOQUEADO'))
               WHERE COALESCE(mp.activo,1)=1
                 AND mp.codigo_mp != ?
                 AND (LOWER(COALESCE(mp.nombre_comercial,'')) LIKE ?
                      OR LOWER(COALESCE(mp.nombre_inci,'')) LIKE ?)
               GROUP BY mp.codigo_mp, mp.nombre_comercial
               HAVING stock_act > 0
               ORDER BY stock_act DESC LIMIT 1""",
            (mid, f'%{nom_norm[:30]}%', f'%{nom_norm[:30]}%'),
        ).fetchone()
        if not candidatos:
            continue
        candidato_id, candidato_nom, candidato_stock = candidatos
        cambios_propuestos.append({
            'huerfano': {'codigo': mid, 'nombre': mnom},
            'reemplazo': {'codigo': candidato_id, 'nombre': candidato_nom,
                          'stock_g': candidato_stock},
        })
        if not dry_run:
            c.execute(
                "UPDATE formula_items SET material_id=?, material_nombre=? "
                "WHERE producto_nombre=? AND material_id=?",
                (candidato_id, candidato_nom, producto, mid),
            )
            aplicados.append({
                'producto': producto,
                'huerfano': mid, 'reemplazo': candidato_id,
            })
    if not dry_run and aplicados:
        try:
            audit_log(c, usuario=u, accion='AUTO_REPARAR_FORMULA',
                      tabla='formula_items', registro_id=producto,
                      despues={'cambios': aplicados})
        except Exception:
            pass
        conn.commit()
    return jsonify({
        'producto': producto,
        'dry_run': dry_run,
        'cambios_propuestos': cambios_propuestos,
        'aplicados': aplicados,
        'mensaje': (
            f'{len(cambios_propuestos)} cambio(s) propuesto(s)'
            if dry_run else f'{len(aplicados)} cambio(s) aplicado(s)'
        ),
    })


@bp.route('/api/produccion/diagnose/<path:producto>', methods=['GET'])
def produccion_diagnose(producto):
    """Sprint Fabricación PRO · 20-may-2026 · Sebastián:
    "necesita Centella extract 7g · dice hay 0g pero en Bodega hay lotes".

    Diagnostica por qué una fórmula no encuentra stock que SÍ existe.
    Causas comunes:
      a) material_id en formula_items quedó huérfano tras unificación
         (UPDATEó movimientos pero NO formula_items)
      b) Los lotes están en cuarentena/rechazados
      c) Lotes sin código de lote (lote='S/L' o NULL)
      d) maestro_mps.activo = 0 para ese material_id

    GET · devuelve por MP: material_id buscado · existe en maestro · stock
    encontrado · lotes similares por nombre (potenciales matches).
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    items = c.execute(
        "SELECT material_id, material_nombre, porcentaje FROM formula_items WHERE producto_nombre=?",
        (producto,),
    ).fetchall()
    if not items:
        return jsonify({'error': f'producto "{producto}" sin fórmula'}), 404
    diag = []
    for mid, mnom, pct in items:
        info = {
            'material_id': mid,
            'material_nombre': mnom,
            'porcentaje': float(pct or 0),
            'problemas': [],
        }
        # 1. Existe en maestro_mps?
        mp_row = c.execute(
            "SELECT nombre_comercial, nombre_inci, COALESCE(activo,1), tipo "
            "FROM maestro_mps WHERE codigo_mp=?", (mid,),
        ).fetchone()
        if not mp_row:
            info['en_maestro'] = False
            info['problemas'].append(f'codigo_mp "{mid}" NO existe en maestro_mps')
        else:
            info['en_maestro'] = True
            info['maestro_nombre'] = mp_row[0]
            info['maestro_inci'] = mp_row[1]
            info['maestro_activo'] = bool(mp_row[2])
            if not mp_row[2]:
                info['problemas'].append(f'maestro_mps.activo=0 para "{mid}"')
        # 2. Stock por lote (sin filtros cuarentena/rechazado)
        try:
            todos = c.execute(
                """SELECT lote,
                          SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock,
                          MAX(COALESCE(estado_lote,''))
                   FROM movimientos
                   WHERE material_id=?
                   GROUP BY lote HAVING stock > 0""",
                (mid,),
            ).fetchall()
        except Exception:
            todos = []
        info['lotes_con_stock'] = len(todos)
        info['stock_total_g'] = sum(float(l[1] or 0) for l in todos)
        info['lotes_detalle'] = [{
            'lote': r[0] or 'S/L',
            'stock_g': float(r[1] or 0),
            'estado_lote': r[2] or '',
        } for r in todos[:10]]
        # 3. Stock DESPUÉS de aplicar filtros del FEFO
        try:
            disp = c.execute(
                """SELECT COUNT(*),
                          COALESCE(SUM(stock_t),0)
                   FROM (SELECT lote,
                                SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_t
                         FROM movimientos
                         WHERE material_id=? AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                           AND (estado_lote IS NULL OR estado_lote NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO'))
                         GROUP BY lote HAVING stock_t > 0)""",
                (mid,),
            ).fetchone()
            info['fefo_disponibles'] = int(disp[0] or 0)
            info['fefo_g'] = float(disp[1] or 0)
        except Exception:
            info['fefo_disponibles'] = 0
            info['fefo_g'] = 0
        # 4. Si stock_total > 0 pero fefo_g = 0 → todo está en cuarentena/rechazado/sin lote
        if info['stock_total_g'] > 0 and info['fefo_g'] == 0:
            info['problemas'].append(
                'Stock existe pero TODO está en cuarentena/rechazado o sin código de lote · '
                'no es usable por FEFO'
            )
        # 5. Si stock_total = 0 · buscar por nombre similar en maestro_mps
        if info['stock_total_g'] == 0 and mnom:
            nom_norm = mnom.strip().lower()[:30]
            similares = c.execute(
                """SELECT codigo_mp, nombre_comercial, COALESCE(nombre_inci,'')
                   FROM maestro_mps
                   WHERE LOWER(COALESCE(nombre_comercial,'')) LIKE ?
                      OR LOWER(COALESCE(nombre_inci,'')) LIKE ?
                   LIMIT 5""",
                (f'%{nom_norm}%', f'%{nom_norm}%'),
            ).fetchall()
            info['mps_similares_por_nombre'] = [{
                'codigo_mp': s[0], 'nombre_comercial': s[1], 'nombre_inci': s[2],
            } for s in similares]
            if similares:
                info['problemas'].append(
                    f'codigo_mp "{mid}" sin stock · pero hay {len(similares)} MPs con '
                    'nombre similar (posible duplicado unificado mal · usar Maestro MP '
                    '→ Unificar duplicados → Actualizar fórmulas)'
                )
        diag.append(info)
    return jsonify({'producto': producto, 'diagnostico': diag})


@bp.route('/api/produccion/<int:pid>/ajustar-cantidad', methods=['POST'])
def produccion_ajustar_cantidad(pid):
    """Sprint Fabricación PRO · 20-may-2026 · Sebastián:
    "registró 29 kg cuando era 30 kg, me ayudas a corregirlo".

    Permite a admin ajustar la cantidad de una producción ya registrada.
    Recalcula el delta de MPs descontadas:
      - Si delta > 0: descontar más MP (FEFO sobre lotes ya usados u otros)
      - Si delta < 0: crear Entrada compensatoria (devolver MP al stock)

    Body: {nueva_cantidad_kg: float, motivo: str (≥10 chars)}
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if (u or '').strip().lower() not in {x.lower() for x in ADMIN_USERS}:
        return jsonify({'error': 'solo admin puede ajustar cantidad de producción'}), 403
    body = request.get_json(silent=True) or {}
    try:
        nueva = float(body.get('nueva_cantidad_kg') or 0)
    except (ValueError, TypeError):
        nueva = 0
    motivo = (body.get('motivo') or '').strip()
    if nueva <= 0:
        return jsonify({'error': 'nueva_cantidad_kg > 0 requerida'}), 400
    if len(motivo) < 10:
        return jsonify({'error': 'motivo ≥10 chars requerido (audit INVIMA)'}), 400
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT producto, cantidad, COALESCE(lote,'') FROM producciones WHERE id=?",
        (pid,),
    ).fetchone()
    if not row:
        return jsonify({'error': 'producción no existe'}), 404
    producto, actual, lote_ref = row[0], float(row[1] or 0), row[2] or f'PROD-{pid:05d}'
    delta_kg = round(nueva - actual, 4)
    if abs(delta_kg) < 0.001:
        return jsonify({'ok': True, 'mensaje': 'Sin cambio (diferencia <0.001kg)',
                        'cantidad_kg': actual})
    delta_g = delta_kg * 1000
    # Obtener fórmula al momento (snapshot si existe, sino actual)
    items = []
    try:
        snap_row = c.execute(
            "SELECT formula_snapshot_json FROM producciones WHERE id=?", (pid,),
        ).fetchone()
        if snap_row and snap_row[0]:
            import json as _json
            items = _json.loads(snap_row[0])
    except Exception:
        items = []
    if not items:
        items_rows = c.execute(
            "SELECT material_id, material_nombre, porcentaje FROM formula_items WHERE producto_nombre=?",
            (producto,),
        ).fetchall()
        items = [{'material_id': r[0], 'material_nombre': r[1], 'porcentaje': r[2]} for r in items_rows]
    if not items:
        return jsonify({'error': f'Producto "{producto}" sin fórmula · no se puede ajustar MPs'}), 400
    from datetime import datetime as _dt
    fecha_ajuste = _dt.now().isoformat()
    movimientos_aplicados = []
    try:
        # MPs ilimitadas (agua, etc.) saltan validación stock
        try:
            from .programacion import _MP_UNLIMITED, _norm_mp_name
            unlimited = {x.upper() for x in _MP_UNLIMITED}
            def _is_unl(n):
                return _norm_mp_name(n or '').upper() in unlimited
        except Exception:
            def _is_unl(n): return False

        if delta_kg > 0:
            # ─ Descontar MP adicional (FEFO) ─
            faltantes = []
            plan = []
            for it in items:
                mid = it['material_id']
                mnom = it['material_nombre']
                pct = float(it['porcentaje'] or 0)
                g_extra = round((pct / 100.0) * delta_g, 2)
                if g_extra <= 0: continue
                if _is_unl(mnom):
                    plan.append({'mid': mid, 'mnom': mnom, 'g': g_extra, 'lote': 'unlimited', 'unlimited': True})
                    continue
                # FEFO
                fefo = c.execute("""SELECT lote,
                    SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock
                    FROM movimientos
                    WHERE material_id=? AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                      AND (estado_lote IS NULL OR estado_lote NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO'))
                    GROUP BY lote HAVING stock > 0
                    ORDER BY COALESCE(NULLIF(CAST(MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) AS TEXT),''),'9999-12-31') ASC""",
                    (mid,)).fetchall()
                disp = sum(float(l[1] or 0) for l in fefo)
                if disp + 0.01 < g_extra:
                    faltantes.append({'material': mnom, 'falta_g': round(g_extra - disp, 2)})
                    continue
                g_rest = g_extra
                for l in fefo:
                    if g_rest <= 0: break
                    g_use = round(min(g_rest, float(l[1])), 2)
                    plan.append({'mid': mid, 'mnom': mnom, 'g': g_use, 'lote': l[0], 'unlimited': False})
                    g_rest = round(g_rest - g_use, 2)
            if faltantes:
                return jsonify({
                    'error': 'Stock insuficiente para ajustar +' + str(delta_kg) + 'kg',
                    'faltantes': faltantes,
                }), 422
            for p in plan:
                c.execute(
                    "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, observaciones, lote, operador) VALUES (?,?,?,?,?,?,?,?)",
                    (p['mid'], p['mnom'], p['g'], 'Salida', fecha_ajuste,
                     f"AJUSTE+:{lote_ref}:{producto} +{delta_kg}kg · " + motivo[:120],
                     None if p['unlimited'] else p['lote'], u),
                )
                movimientos_aplicados.append({'material': p['mnom'], 'g': p['g'], 'tipo': 'Salida extra'})
        else:
            # ─ Devolver MP (delta negativo) · Entrada compensatoria ─
            # Buscar movimientos de Salida originales del lote para
            # devolver al mismo lote MP de procedencia.
            for it in items:
                mid = it['material_id']
                mnom = it['material_nombre']
                pct = float(it['porcentaje'] or 0)
                g_dev = round((pct / 100.0) * abs(delta_g), 2)
                if g_dev <= 0: continue
                # Buscar la salida original más grande de este MP en esta producción
                orig = c.execute(
                    """SELECT lote, SUM(cantidad)
                       FROM movimientos
                       WHERE material_id=? AND tipo='Salida'
                         AND observaciones LIKE ?
                       GROUP BY lote
                       ORDER BY SUM(cantidad) DESC LIMIT 1""",
                    (mid, f'%FEFO:{lote_ref}:%'),
                ).fetchone()
                lote_dev = orig[0] if orig and orig[0] else None
                if _is_unl(mnom): lote_dev = None
                c.execute(
                    "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, observaciones, lote, operador) VALUES (?,?,?,?,?,?,?,?)",
                    (mid, mnom, g_dev, 'Entrada', fecha_ajuste,
                     f"AJUSTE-:{lote_ref}:{producto} {delta_kg}kg (devolución) · " + motivo[:120],
                     lote_dev, u),
                )
                movimientos_aplicados.append({'material': mnom, 'g': g_dev, 'tipo': 'Entrada devolución'})
        # Update cantidad
        c.execute("UPDATE producciones SET cantidad=? WHERE id=?", (nueva, pid))
        # Audit
        try:
            audit_log(c, usuario=u, accion='AJUSTAR_CANTIDAD_PRODUCCION',
                      tabla='producciones', registro_id=pid,
                      antes={'cantidad_kg': actual},
                      despues={'cantidad_kg': nueva, 'delta_kg': delta_kg,
                               'motivo': motivo[:200],
                               'movs': len(movimientos_aplicados)})
        except Exception:
            pass
        conn.commit()
        return jsonify({
            'ok': True,
            'pid': pid, 'lote': lote_ref,
            'cantidad_anterior_kg': actual,
            'cantidad_nueva_kg': nueva,
            'delta_kg': delta_kg,
            'movimientos_aplicados': movimientos_aplicados,
            'mensaje': f'Producción {lote_ref} ajustada de {actual}kg a {nueva}kg ({"+" if delta_kg>0 else ""}{delta_kg}kg)',
        })
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        __import__('logging').getLogger('inventario').exception(
            'ajustar_cantidad_produccion · pid=%s err=%s', pid, e)
        return jsonify({'error': f'Falla al ajustar: {e}', 'tipo': type(e).__name__}), 500


@bp.route('/api/produccion/<int:pid>/rotulo-reimprimir', methods=['GET'])
def produccion_rotulo_reimprimir(pid):
    """Sprint Fabricación PRO · genera HTML imprimible con los rótulos
    del lote PT · sirve para reimprimir si se perdieron los originales."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    h = c.execute(
        """SELECT id, producto, cantidad, fecha, COALESCE(presentacion,''),
                  COALESCE(lote,''), COALESCE(operador,'')
           FROM producciones WHERE id=?""",
        (pid,),
    ).fetchone()
    if not h:
        return jsonify({'error': 'producción no existe'}), 404
    lote_ref = h[5] or f'PROD-{h[0]:05d}'
    presentacion = h[4] or '—'
    fecha_corta = (h[3] or '')[:10]
    def _esc(s):
        return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Rótulos · {_esc(lote_ref)}</title>
<style>
@page {{ size: A4; margin: 8mm; }}
body {{ font-family: Arial, sans-serif; margin: 0; padding: 10px; }}
h1 {{ margin: 0 0 14px; color: #0f766e; font-size: 18px; }}
.rot {{ display: inline-block; width: 90mm; height: 50mm; border: 1px dashed #94a3b8; padding: 6mm; margin: 2mm; vertical-align: top; box-sizing: border-box; }}
.rot b {{ font-size: 14px; color: #0f172a; }}
.rot .meta {{ font-size: 10px; color: #475569; margin-top: 4px; line-height: 1.4; }}
.rot .lote {{ font-size: 18px; font-weight: 800; color: #dc2626; margin-top: 6px; letter-spacing: 1px; }}
@media print {{ button {{ display: none; }} }}
</style></head><body>
<button onclick="window.print()" style="background:#0f766e;color:#fff;padding:10px 20px;border:none;border-radius:5px;cursor:pointer">🖨 Imprimir</button>
<h1>Rótulos · {_esc(lote_ref)}</h1>"""
    for i in range(6):
        html += f"""
<div class="rot">
  <b>Espagiria Laboratorio · ÁNIMUS Lab</b><br>
  <span style="font-size:13px;font-weight:700">{_esc(h[1] or '—')}</span>
  <div class="lote">LOTE {_esc(lote_ref)}</div>
  <div class="meta">
    Presentación: {_esc(presentacion)}<br>
    Cantidad: {h[2]} kg total · Fab: {fecha_corta}<br>
    Operario: {_esc(h[6])}<br>
    Hecho en Colombia · INVIMA cosmético
  </div>
</div>"""
    html += '</body></html>'
    from flask import Response as _R
    return _R(html, mimetype='text/html; charset=utf-8')

@bp.route('/api/produccion/simular', methods=['POST'])
def simular_produccion():
    """Pre-check de stock FEFO + estimado de costo sin commitear ningun movimiento."""
    data = request.get_json(silent=True) or {}
    producto = data.get('producto', '')
    try:
        cantidad_kg = float(data.get('cantidad_kg', 1))
    except (TypeError, ValueError):
        return jsonify({'error': 'cantidad_kg inválida', 'factible': False}), 400
    cantidad_g = cantidad_kg * 1000
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT fi.material_id, fi.material_nombre, fi.porcentaje,
                        COALESCE(m.precio_referencia, 0),
                        COALESCE(m.controla_stock, 1)
                 FROM formula_items fi
                 LEFT JOIN maestro_mps m ON fi.material_id = m.codigo_mp
                 WHERE fi.producto_nombre=? AND fi.producto_nombre NOT IN
                       (SELECT producto_nombre FROM formula_headers WHERE COALESCE(activo,1)=0)""", (producto,))
    items = c.fetchall()
    if not items:
        return jsonify({'error': f'Formula no encontrada o descontinuada: {producto}', 'factible': False}), 404
    resultado = []
    factible = True
    costo_total = 0.0
    sin_precio = 0
    # Audit 4-jun · MISMA resolución que el descuento real (M1): resolver el código
    # de fórmula al de bodega y excluir los mismos 6 estados → "Verificar Stock"
    # deja de mentir 0g cuando el stock está bajo un código duplicado/otro nombre.
    try:
        from blueprints.programacion import (_resolver_material_bodega as _resolver_mp,
                                             _ESTADOS_LOTE_NO_PRODUCIBLES as _NP6)
    except Exception:
        _resolver_mp, _NP6 = None, None
    for mat_id, mat_nombre, pct, precio_kg, controla_stock in items:
        g_req = round((pct / 100) * cantidad_g, 2)
        # MP infinita / fabricada en casa (AGUA del lab) → no se controla stock:
        # siempre suficiente, nunca faltante, no bloquea la producción.
        if int(controla_stock or 0) == 0:
            resultado.append({
                'material_id': mat_id, 'material_nombre': mat_nombre,
                'porcentaje': pct, 'g_requerido': g_req,
                'g_disponible': g_req, 'g_faltante': 0,
                'suficiente': True, 'no_controla_stock': True,
                'precio_kg': round(precio_kg or 0, 2), 'costo': 0,
            })
            continue
        if _resolver_mp and _NP6:
            try:
                cod_bodega = _resolver_mp(c, mat_id, mat_nombre) or mat_id
            except Exception:
                cod_bodega = mat_id
            # M-1 (Sebastian 12-jun): alinear "Verificar Stock" con la seleccion REAL
            # de FEFO -> solo lotes con lote real (no S/L) y stock>0.01 por lote,
            # excluyendo no-producibles (incl BLOQUEADO). Antes era un SUM plano que
            # sobre-reportaba (sumaba S/L, polvo <0.01 y estados retenidos) -> simular
            # decia factible y el descuento real fallaba con 422.
            _ph = ','.join(['?'] * len(_NP6))
            # M-1 + vencimiento (12-jun): excluir tambien lotes vencidos POR FECHA
            # (fv Entrada < hoy Colombia) aunque el cron diario aun no los marque
            # VENCIDO — mismo limite que job_marcar_vencidos. Asi "Verificar Stock"
            # no dice "alcanza" con un lote que el FEFO real va a rechazar (M5).
            c.execute(f"""SELECT COALESCE(SUM(stk),0) FROM (
                            SELECT SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stk,
                                   MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) AS fv_real
                            FROM movimientos WHERE material_id=?
                              AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                              AND UPPER(COALESCE(estado_lote,'')) NOT IN ({_ph})
                            GROUP BY lote HAVING stk > 0.01
                              AND (fv_real IS NULL OR TRIM(CAST(fv_real AS TEXT))='' OR date(fv_real) >= date('now', '-5 hours')))""",
                      (cod_bodega,) + tuple(_NP6))
        else:
            c.execute("""SELECT COALESCE(SUM(stk),0) FROM (
                            SELECT SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stk,
                                   MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) AS fv_real
                            FROM movimientos WHERE material_id=?
                              AND lote IS NOT NULL AND lote!='' AND lote!='S/L'
                              AND (estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO'))
                            GROUP BY lote HAVING stk > 0.01
                              AND (fv_real IS NULL OR TRIM(CAST(fv_real AS TEXT))='' OR date(fv_real) >= date('now', '-5 hours')))""",
                      (mat_id,))
        g_disp = round(c.fetchone()[0] or 0, 2)
        suf = g_disp >= g_req
        if not suf:
            factible = False
        precio_g = (precio_kg or 0) / 1000.0
        costo_item = round(g_req * precio_g, 2)
        costo_total += costo_item
        if not precio_kg or precio_kg == 0:
            sin_precio += 1
        resultado.append({
            'material_id': mat_id, 'material_nombre': mat_nombre,
            'porcentaje': pct, 'g_requerido': g_req,
            'g_disponible': g_disp,
            'g_faltante': max(0, round(g_req - g_disp, 2)),
            'suficiente': suf,
            'precio_kg': round(precio_kg or 0, 2),
            'costo': costo_item
        })
    faltantes = sum(1 for r in resultado if not r['suficiente'])
    n = len(resultado)
    return jsonify({
        'producto': producto, 'cantidad_kg': cantidad_kg,
        'factible': factible, 'faltantes': faltantes,
        'costo_total': round(costo_total, 2),
        'costo_por_kg': round(costo_total / cantidad_kg, 2) if cantidad_kg > 0 else 0,
        'ingredientes_sin_precio': sin_precio,
        'cobertura_precio_pct': round((n - sin_precio) / n * 100, 1) if n > 0 else 0,
        'ingredientes': sorted(resultado, key=lambda x: x['suficiente'])
    })

def _get_formula_pin():
    """Sprint Fórmulas PRO · 20-may-2026 · PIN runtime con override en
    app_settings. Si admin lo cambia desde UI, queda en BD (no requiere
    setear env var en Render). Sino, fallback al env FORMULA_PIN.
    Defensivo: si la tabla no existe (mig 147 no aplicó aún), no rompe.
    """
    try:
        conn = get_db()
        # Crear tabla si no existe · defensivo en producción
        try:
            conn.execute("""CREATE TABLE IF NOT EXISTS app_settings (
                clave TEXT PRIMARY KEY,
                valor TEXT NOT NULL,
                descripcion TEXT,
                actualizado_at_utc TEXT,
                actualizado_por TEXT,
                tenant_id INTEGER DEFAULT 1
            )""")
        except Exception:
            pass
        row = conn.execute(
            "SELECT valor FROM app_settings WHERE clave='formula_pin' LIMIT 1",
        ).fetchone()
        if row and row[0]:
            return str(row[0])
    except Exception:
        pass
    from config import FORMULA_PIN
    return FORMULA_PIN


@bp.route('/api/formulas/unlock', methods=['POST'])
def formulas_unlock():
    data = request.get_json() or {}
    pin = str(data.get('pin', ''))
    if pin and pin == _get_formula_pin():
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'PIN incorrecto'}), 403


def _ensure_app_settings_table(c):
    """Crea app_settings si no existe (defensivo · si mig 147 no se aplicó
    en producción, esto la crea al primer acceso). Idempotente."""
    try:
        c.execute("""CREATE TABLE IF NOT EXISTS app_settings (
            clave TEXT PRIMARY KEY,
            valor TEXT NOT NULL,
            descripcion TEXT,
            actualizado_at_utc TEXT,
            actualizado_por TEXT,
            tenant_id INTEGER DEFAULT 1
        )""")
    except Exception:
        pass  # tabla ya existe o BD no acepta · sigue


@bp.route('/api/admin/formulas/pin', methods=['GET', 'POST'])
def admin_formulas_pin():
    """Sprint Fórmulas PRO · admin cambia PIN desde UI (sin tocar Render).

    GET  · devuelve estado actual sin revelar valor.
    POST · body {nuevo_pin} · solo admin · 4-32 chars.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    # Case-insensitive admin check · defensivo si la sesión guardó
    # 'Sebastian' o 'SEBASTIAN' por cualquier razón.
    u_norm = (u or '').strip().lower()
    admin_norm = {x.lower() for x in ADMIN_USERS}
    if u_norm not in admin_norm:
        __import__('logging').getLogger('inventario').warning(
            'admin_formulas_pin · acceso denegado · user=%r admin_set=%r',
            u, list(ADMIN_USERS))
        return jsonify({
            'error': f'solo admin · sesión actual: "{u}"',
            'admin_users_esperados': sorted(ADMIN_USERS),
        }), 403
    try:
        conn = get_db(); c = conn.cursor()
    except Exception as _e:
        __import__('logging').getLogger('inventario').error(
            'admin_formulas_pin · get_db falló: %s', _e)
        return jsonify({'error': f'BD inaccesible: {_e}'}), 500
    _ensure_app_settings_table(c)

    try:
        from config import _FORMULA_PIN_INSECURE
        es_random = bool(_FORMULA_PIN_INSECURE)
    except Exception:
        es_random = False

    if request.method == 'GET':
        try:
            row = c.execute(
                "SELECT actualizado_at_utc, actualizado_por FROM app_settings "
                "WHERE clave='formula_pin' LIMIT 1",
            ).fetchone()
        except Exception as _e:
            __import__('logging').getLogger('inventario').warning(
                'admin_formulas_pin GET · query falló: %s', _e)
            row = None
        return jsonify({
            'configurado_en_bd': bool(row),
            'configurado_en_env': not es_random and not bool(row),
            'es_pin_random_efimero': es_random and not bool(row),
            'ultima_actualizacion': (row[0] if row else None),
            'cambiado_por': (row[1] if row else None),
        })

    body = request.get_json(silent=True) or {}
    nuevo = str(body.get('nuevo_pin') or '').strip()
    if len(nuevo) < 4:
        return jsonify({'error': 'PIN debe tener ≥4 caracteres'}), 400
    if len(nuevo) > 32:
        return jsonify({'error': 'PIN máximo 32 caracteres'}), 400
    # Upsert manual · compatible SQLite y PostgreSQL · sin ON CONFLICT
    # con expresiones potencialmente problemáticas.
    from datetime import datetime as _dt
    now_utc = _dt.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    try:
        existe = c.execute(
            "SELECT 1 FROM app_settings WHERE clave='formula_pin' LIMIT 1",
        ).fetchone()
        if existe:
            c.execute(
                "UPDATE app_settings SET valor=?, actualizado_at_utc=?, "
                "actualizado_por=? WHERE clave='formula_pin'",
                (nuevo, now_utc, u),
            )
        else:
            c.execute(
                "INSERT INTO app_settings (clave, valor, descripcion, "
                "actualizado_at_utc, actualizado_por) VALUES "
                "('formula_pin', ?, 'PIN runtime fórmulas · override de env', ?, ?)",
                (nuevo, now_utc, u),
            )
    except Exception as _e:
        __import__('logging').getLogger('inventario').error(
            'admin_formulas_pin POST · upsert falló: %s', _e)
        return jsonify({'error': f'No se pudo guardar el PIN: {_e}',
                        'detalle': 'Revisar logs del servidor'}), 500
    try:
        audit_log(c, usuario=u, accion='FORMULA_PIN_CAMBIADO',
                  tabla='app_settings', registro_id='formula_pin',
                  despues={'len': len(nuevo)})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'mensaje': f'PIN actualizado por {u}'})

@bp.route('/api/formula/costo', methods=['POST'])
def calcular_costo_formula():
    """Calcula costo estimado de un batch sin verificar stock."""
    data = request.get_json(silent=True) or {}
    producto = data.get('producto', '')
    try:
        cantidad_kg = float(data.get('cantidad_kg', 1))
    except (TypeError, ValueError):
        return jsonify({'error': 'cantidad_kg inválida'}), 400
    cantidad_g = cantidad_kg * 1000
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT fi.material_id, fi.material_nombre, fi.porcentaje,
                        COALESCE(m.precio_referencia, 0)
                 FROM formula_items fi
                 LEFT JOIN maestro_mps m ON fi.material_id = m.codigo_mp
                 WHERE fi.producto_nombre=?""", (producto,))
    items = c.fetchall()
    if not items:
        return jsonify({'error': f'Formula no encontrada: {producto}'}), 404
    resultado = []
    costo_total = 0.0
    sin_precio = 0
    for mat_id, mat_nombre, pct, precio_kg in items:
        g_req = round((pct / 100) * cantidad_g, 2)
        precio_g = (precio_kg or 0) / 1000.0
        costo_item = round(g_req * precio_g, 2)
        costo_total += costo_item
        if not precio_kg or precio_kg == 0:
            sin_precio += 1
        resultado.append({
            'material_id': mat_id, 'material_nombre': mat_nombre,
            'porcentaje': pct, 'g_requerido': g_req,
            'precio_kg': round(precio_kg or 0, 2),
            'precio_g': round(precio_g, 5),
            'costo': costo_item
        })
    n = len(resultado)
    return jsonify({
        'producto': producto, 'cantidad_kg': cantidad_kg,
        'costo_total': round(costo_total, 2),
        'costo_por_kg': round(costo_total / cantidad_kg, 2) if cantidad_kg > 0 else 0,
        'ingredientes_sin_precio': sin_precio,
        'cobertura_precio_pct': round((n - sin_precio) / n * 100, 1) if n > 0 else 0,
        'ingredientes': sorted(resultado, key=lambda x: x['costo'], reverse=True)
    })

@bp.route('/api/trazabilidad/lote-pt/<lote_ref>')
def trazabilidad_lote_pt(lote_ref):
    """Traza hacia atrás: dado un lote PT (PROD-00001) devuelve MPs consumidas, proveedor, fecha vencimiento."""
    conn = get_db(); c = conn.cursor()
    # Producción base
    c.execute("SELECT id, producto, cantidad, fecha, operador, observaciones FROM producciones WHERE lote=? OR id=?",
              (lote_ref, lote_ref.replace('PROD-','').lstrip('0') or 0))
    prod = c.fetchone()
    if not prod:
        return jsonify({'error': f'Lote no encontrado: {lote_ref}', 'lote_ref': lote_ref}), 404
    prod_data = {'id': prod[0], 'producto': prod[1], 'cantidad_kg': prod[2],
                 'fecha': prod[3], 'operador': prod[4] or '', 'observaciones': prod[5] or ''}
    # MPs consumidas — buscar Salidas etiquetadas con este lote_ref O por fecha+producto (legacy)
    c.execute("""SELECT material_id, material_nombre, SUM(cantidad) as g_total,
                        GROUP_CONCAT(DISTINCT lote) as lotes_mp,
                        GROUP_CONCAT(DISTINCT proveedor) as proveedores
                 FROM movimientos
                 WHERE tipo='Salida'
                   AND (observaciones LIKE ? OR (fecha=? AND observaciones LIKE ?))
                 GROUP BY material_id, material_nombre
                 ORDER BY material_nombre""",
              (f'FEFO:{lote_ref}:%', prod[3], f'%{prod[1]}%'))
    mps = [{'material_id': r[0], 'material_nombre': r[1], 'g_consumido': round(r[2], 2),
             'lotes_mp': [l for l in (r[3] or '').split(',') if l and l != 'None'],
             'proveedores': list(set([p for p in (r[4] or '').split(',') if p and p != 'None']))}
           for r in c.fetchall()]
    # Para cada lote de MP consumido, obtener info del ingreso original
    detalle_lotes = []
    for mp in mps:
        for lote_mp in mp['lotes_mp']:
            c.execute("""SELECT fecha, proveedor, numero_oc, numero_factura, fecha_vencimiento, estado_lote
                         FROM movimientos WHERE lote=? AND tipo='Entrada' AND material_id=?
                         ORDER BY fecha DESC LIMIT 1""", (lote_mp, mp['material_id']))
            row = c.fetchone()
            if row:
                detalle_lotes.append({
                    'material_id': mp['material_id'], 'material_nombre': mp['material_nombre'],
                    'lote_mp': lote_mp, 'fecha_ingreso': row[0][:10] if row[0] else '',
                    'proveedor': row[1] or '', 'numero_oc': row[2] or '',
                    'numero_factura': row[3] or '', 'fecha_vencimiento': row[4] or '',
                    'estado_lote': row[5] or 'VIGENTE'
                })
    # Despachos que usaron este PT batch
    c.execute("""SELECT d.numero, cl.nombre, d.fecha,
                        di.sku, di.descripcion, di.cantidad
                 FROM despachos_items di
                 JOIN despachos d ON di.numero_despacho=d.numero
                 LEFT JOIN clientes cl ON d.cliente_id=cl.id
                 WHERE di.lote_pt=? OR di.lote_pt LIKE ?""", (lote_ref, f'%{lote_ref}%'))
    despachos = [{'numero': r[0], 'cliente': r[1] or '', 'fecha': r[2], 'sku': r[3],
                  'descripcion': r[4], 'cantidad': r[5]} for r in c.fetchall()]
    return jsonify({
        'lote_ref': lote_ref, 'produccion': prod_data,
        'mps_consumidas': mps, 'detalle_lotes_mp': detalle_lotes,
        'despachos': despachos,
        'trazabilidad_completa': len(mps) > 0
    })

@bp.route('/api/trazabilidad/lote-mp/<path:lote_mp>')
def trazabilidad_lote_mp(lote_mp):
    """Traza hacia adelante: dado un lote de MP devuelve en qué producciones se usó y a qué clientes llegó."""
    conn = get_db(); c = conn.cursor()
    # Ingreso del lote
    c.execute("""SELECT material_id, material_nombre, cantidad, fecha, proveedor,
                        numero_oc, numero_factura, fecha_vencimiento, estado_lote
                 FROM movimientos WHERE lote=? AND tipo='Entrada' LIMIT 1""", (lote_mp,))
    ingreso = c.fetchone()
    if not ingreso:
        return jsonify({'error': f'Lote MP no encontrado: {lote_mp}'}), 404
    mat_info = {
        'material_id': ingreso[0], 'material_nombre': ingreso[1],
        'cantidad_kg_ingresada': round((ingreso[2] or 0) / 1000, 3),
        'fecha_ingreso': ingreso[3][:10] if ingreso[3] else '', 'proveedor': ingreso[4] or '',
        'numero_oc': ingreso[5] or '', 'numero_factura': ingreso[6] or '',
        'fecha_vencimiento': ingreso[7] or '', 'estado_lote': ingreso[8] or 'VIGENTE'
    }
    # Salidas — producciones que consumieron este lote
    c.execute("""SELECT observaciones, cantidad, fecha FROM movimientos
                 WHERE lote=? AND tipo='Salida' ORDER BY fecha""", (lote_mp,))
    salidas = c.fetchall()
    producciones_ref = []
    for obs, cant, fec in salidas:
        # obs format: "FEFO:PROD-00001:Suero TRX x 10kg" or legacy "FEFO: Suero TRX x 10kg"
        lote_prod = ''
        if obs and obs.startswith('FEFO:PROD-'):
            parts = obs.split(':')
            lote_prod = parts[1] if len(parts) > 1 else ''
        producciones_ref.append({
            'lote_produccion': lote_prod, 'g_consumido': round(cant, 2),
            'fecha': fec[:10] if fec else '', 'observaciones': obs or ''
        })
    # Detallar producciones únicas
    lotes_prod_unicos = list(set(p['lote_produccion'] for p in producciones_ref if p['lote_produccion']))
    producciones_detalle = []
    for lp in lotes_prod_unicos:
        c.execute("SELECT producto, cantidad, fecha, operador FROM producciones WHERE lote=?", (lp,))
        pr = c.fetchone()
        if pr:
            # Despachos desde este lote PT
            c.execute("""SELECT d.numero, cl.nombre, d.fecha, di.cantidad
                         FROM despachos_items di
                         JOIN despachos d ON di.numero_despacho=d.numero
                         LEFT JOIN clientes cl ON d.cliente_id=cl.id
                         WHERE di.lote_pt=?""", (lp,))
            dsps = [{'numero': r[0], 'cliente': r[1] or '', 'fecha': r[2], 'cantidad': r[3]} for r in c.fetchall()]
            producciones_detalle.append({
                'lote_ref': lp, 'producto': pr[0], 'cantidad_kg': pr[1],
                'fecha': pr[2][:10] if pr[2] else '', 'operador': pr[3] or '',
                'despachos': dsps
            })
    return jsonify({
        'lote_mp': lote_mp, 'material': mat_info,
        'salidas': producciones_ref,
        'producciones': producciones_detalle,
        'clientes_afectados': list(set(
            d['cliente'] for p in producciones_detalle for d in p['despachos']
        ))
    })

@bp.route('/api/analisis-abc')
def get_analisis_abc():
    """Análisis ABC Pareto · Sebastián 20-may-2026 · 3 modos + filtros.

    Query params:
      modo:
        valor (default) · ordena por stock × precio_referencia (Pareto financiero)
        consumo_90d · ordena por consumo proyectado (salidas últimos 90d × precio)
        consumo_180d / consumo_365d
        stock_actual · ordena por gramos en bodega (modo viejo, compat)
      excluir_cuarentena: 1 (default 0)
      subtipo: filtra por maestro_mps.tipo (Activo, Emoliente, ...) opcional
      incluir_sin_movimientos: 1 (default 1 en modo valor, 0 en modos consumo)
      tipo_material: MP (default) | MEE

    Devuelve cada MP con:
      material_id, nombre_comercial, nombre_inci, proveedor, subtipo,
      origen (china/colombia/desconocido), stock_g, precio_kg, valor_cop,
      consumo_90d_g, metric (el valor usado para Pareto · depende del modo),
      pct_acumulado, ranking, clasificacion (A/B/C)
    """
    u, err, code = _require_session()
    if err:
        return err, code
    modo = (request.args.get('modo') or 'valor').strip().lower()
    excluir_cuarentena = (request.args.get('excluir_cuarentena') or '').strip() in ('1', 'true', 'yes')
    subtipo_filtro = (request.args.get('subtipo') or '').strip()
    tipo_material = (request.args.get('tipo_material') or 'MP').strip().upper()
    incluir_sin_mov_raw = request.args.get('incluir_sin_movimientos')
    if incluir_sin_mov_raw is None:
        incluir_sin_mov = (modo == 'valor' or modo == 'stock_actual')
    else:
        incluir_sin_mov = incluir_sin_mov_raw in ('1', 'true', 'yes')

    # Determinar ventana de consumo si modo lo requiere
    ventana_dias = None
    if modo.startswith('consumo_'):
        try:
            ventana_dias = int(modo.split('_')[1].rstrip('d'))
        except (ValueError, IndexError):
            ventana_dias = 90

    conn = get_db()
    c = conn.cursor()

    if tipo_material == 'MEE':
        # Camino simple para MEE: usa maestro_mee directamente
        # stock CANÓNICO desde SUM(movimientos_mee) (no el cache stock_actual que
        # driftea · igual que _mee_stock_real). Columna real = mee_codigo.
        rows = c.execute(
            """SELECT m.codigo, m.descripcion, '' as nombre_inci, COALESCE(m.proveedor,''),
                      COALESCE(m.categoria,''), '' as origen,
                      COALESCE((SELECT SUM(CASE WHEN mm.tipo='Entrada' THEN mm.cantidad
                                                WHEN mm.tipo='Salida'  THEN -mm.cantidad
                                                WHEN mm.tipo='Ajuste'  THEN mm.cantidad
                                                ELSE 0 END)
                                FROM movimientos_mee mm
                                WHERE mm.mee_codigo=m.codigo AND COALESCE(mm.anulado,0)=0), m.stock_actual, 0) as stock,
                      0 as precio  -- maestro_mee no tiene columna de precio (ABC MEE = por gramos/consumo, no valor)
               FROM maestro_mee m
               WHERE COALESCE(m.estado, 'Activo') = 'Activo'
               """,
        ).fetchall()
        consumo_map = {}
        if ventana_dias:
            for r in c.execute(
                f"""SELECT mee_codigo, COALESCE(SUM(cantidad), 0)
                    FROM movimientos_mee
                    WHERE tipo IN ('Salida','salida','SALIDA') AND COALESCE(anulado,0)=0
                      AND fecha >= date('now', '-5 hours', '-{int(ventana_dias)} days')
                    GROUP BY mee_codigo""",
            ).fetchall():
                consumo_map[r[0]] = float(r[1] or 0)
        items_raw = []
        for r in rows:
            cod, nom, inci, prov, subt, orig, stock, precio = r
            stock = max(float(stock or 0), 0)
            precio = float(precio or 0)
            valor_cop = stock * precio
            consumo = consumo_map.get(cod, 0.0)
            items_raw.append({
                'material_id': cod or '', 'nombre_comercial': nom or '',
                'nombre_inci': inci, 'proveedor': prov,
                'subtipo': subt, 'origen': orig,
                'stock_g': round(stock, 2), 'precio_kg': round(precio, 2),
                'valor_cop': round(valor_cop, 2),
                'consumo_90d_g': round(consumo, 2),
            })
    else:
        # MP path · LEFT JOIN desde maestro_mps + movimientos agregados
        cuarentena_filter = ""
        if excluir_cuarentena:
            cuarentena_filter = " AND UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO')"
        # Stock agregado por material_id excluyendo cuarentena si aplica
        rows_stock = c.execute(
            f"""SELECT material_id,
                       COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                                          WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                                          ELSE 0 END), 0) as stock
                FROM movimientos
                WHERE material_id IS NOT NULL AND material_id != ''
                  {cuarentena_filter}
                GROUP BY material_id""",
        ).fetchall()
        stock_map = {r[0]: max(float(r[1] or 0), 0) for r in rows_stock}
        # Consumo en ventana si aplica
        consumo_map = {}
        if ventana_dias:
            for r in c.execute(
                f"""SELECT material_id, COALESCE(SUM(cantidad), 0)
                    FROM movimientos
                    WHERE tipo IN ('Salida','salida','SALIDA')
                      AND fecha >= date('now', '-5 hours', '-{int(ventana_dias)} days')
                      AND material_id IS NOT NULL AND material_id != ''
                    GROUP BY material_id""",
            ).fetchall():
                consumo_map[r[0]] = float(r[1] or 0)
        # Catálogo de MPs activas (LEFT JOIN para incluir las sin mov si flag)
        try:
            mps = c.execute(
                """SELECT codigo_mp, COALESCE(nombre_comercial,''),
                          COALESCE(nombre_inci,''), COALESCE(proveedor,''),
                          COALESCE(tipo,''), COALESCE(precio_referencia, 0)
                   FROM maestro_mps
                   WHERE COALESCE(activo, 1) = 1
                     AND UPPER(COALESCE(tipo_material,'MP')) = 'MP'""",
            ).fetchall()
        except Exception:
            mps = []
        # Origen heurístico por proveedor (china/colombia/desconocido)
        def _origen(prov):
            p = (prov or '').lower()
            CHINA_HINTS = ['lyphar', 'inchemical', 'guanfu', 'guangzhou',
                            'china', 'shanghai', 'xian', 'wuhan']
            LOCAL_HINTS = ['quimica', 'colombia', 'cosmo', 'andina', 'bogot',
                            'medell', 'cali']
            if any(h in p for h in CHINA_HINTS):
                return 'china'
            if any(h in p for h in LOCAL_HINTS):
                return 'colombia'
            return 'desconocido' if not p else 'otro'

        items_raw = []
        for mp in mps:
            cod, nom, inci, prov, subt, precio_ref = mp
            stock = stock_map.get(cod, 0.0)
            consumo = consumo_map.get(cod, 0.0)
            precio = float(precio_ref or 0)
            valor_cop = stock * precio
            # Filtros
            if subtipo_filtro and (subt or '').lower() != subtipo_filtro.lower():
                continue
            if not incluir_sin_mov:
                # Si modo es consumo · excluir items sin consumo en la ventana.
                # Si modo es stock · excluir items sin stock.
                # Si modo es valor · excluir solo si stock=0 Y consumo=0.
                if modo.startswith('consumo_') and consumo <= 0:
                    continue
                elif modo == 'stock_actual' and stock <= 0:
                    continue
                elif modo == 'valor' and stock <= 0 and consumo <= 0:
                    continue
            items_raw.append({
                'material_id': cod or '', 'nombre_comercial': nom,
                'nombre_inci': inci, 'proveedor': prov,
                'subtipo': subt, 'origen': _origen(prov),
                'stock_g': round(stock, 2), 'precio_kg': round(precio, 2),
                'valor_cop': round(valor_cop, 2),
                'consumo_90d_g': round(consumo, 2),
            })

    # Elegir métrica según modo
    def _metric(item):
        if modo == 'valor':
            return item['valor_cop']
        if modo == 'stock_actual':
            return item['stock_g']
        # consumo_*
        return item['consumo_90d_g'] * (item['precio_kg'] or 1) if modo.startswith('consumo_') else item['valor_cop']

    items_pos = [{**it, 'metric': _metric(it)} for it in items_raw]
    items_pos.sort(key=lambda x: -x['metric'])

    total = sum(it['metric'] for it in items_pos if it['metric'] > 0)
    if total <= 0:
        # Modo "valor" pero sin precios cargados · fallback a stock_actual
        if modo == 'valor':
            for it in items_pos:
                it['metric'] = it['stock_g']
            items_pos.sort(key=lambda x: -x['metric'])
            total = sum(it['metric'] for it in items_pos if it['metric'] > 0)

    cumulative = 0.0
    abc_items = []
    for idx, it in enumerate(items_pos, start=1):
        m = it['metric']
        prev_pct = (cumulative / total * 100) if total > 0 else 0
        cumulative += m
        pct = (cumulative / total * 100) if total > 0 else 0
        if m <= 0:
            clasif = 'D'  # sin métrica · no califica Pareto
        else:
            clasif = 'A' if prev_pct < 80 else ('B' if prev_pct < 95 else 'C')
        abc_items.append({
            **it,
            'metric': round(m, 2),
            'pct_acumulado': round(pct, 2),
            'ranking': idx,
            'clasificacion': clasif,
        })

    # Stats summary
    counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
    valor_por_clase = {'A': 0.0, 'B': 0.0, 'C': 0.0, 'D': 0.0}
    for it in abc_items:
        counts[it['clasificacion']] = counts.get(it['clasificacion'], 0) + 1
        valor_por_clase[it['clasificacion']] = valor_por_clase.get(it['clasificacion'], 0) + it['metric']

    return jsonify({
        'modo': modo,
        'tipo_material': tipo_material,
        'excluir_cuarentena': excluir_cuarentena,
        'subtipo_filtro': subtipo_filtro,
        'incluir_sin_movimientos': incluir_sin_mov,
        'ventana_consumo_dias': ventana_dias,
        'total_items': len(abc_items),
        'total_metric': round(total, 2),
        'metric_unit': 'COP' if modo in ('valor',) else (
            'COP estimado' if modo.startswith('consumo_') else 'g'),
        'counts': counts,
        'valor_por_clase': {k: round(v, 2) for k, v in valor_por_clase.items()},
        'items': abc_items,
        # Compat con UI vieja
        'items_legacy': [
            {'material': it['nombre_comercial'] or it['material_id'],
             'cantidad': it['stock_g'], 'valor': f'{it["pct_acumulado"]:.1f}%',
             'clasificacion': it['clasificacion']}
            for it in abc_items
        ],
    })

@bp.route('/api/alertas', methods=['GET', 'POST'])
def handle_alertas():
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        mat_id = data.get('material_id')
        if not mat_id:
            return jsonify({'error': 'material_id requerido'}), 400
        c.execute('INSERT INTO alertas (material_id, material_nombre, stock_actual, stock_minimo, fecha, estado) VALUES (?,?,?,?,?,?)',
                  (mat_id, data.get('material_nombre',''), data.get('stock_actual',0),
                   data.get('stock_minimo',0), datetime.now().isoformat(), 'Activa'))
        conn.commit()
        return jsonify({'message': 'Alerta creada'}), 201
    c.execute('SELECT material_nombre, stock_actual, stock_minimo, estado, fecha FROM alertas ORDER BY fecha DESC')
    alertas = [{'material_nombre': r[0], 'stock_actual': r[1], 'stock_minimo': r[2], 'estado': r[3], 'fecha': r[4]} for r in c.fetchall()]
    return jsonify({'alertas': alertas})

# ────────────────────────────────────────────────────────────────────
# Sprint Alertas PRO · Sebastián 20-may-2026 · endpoint consolidado
# Agrupa todas las alertas (sin stock / bajo mínimo / vencidos /
# próximos / cuarentena / MEE) en una sola llamada con stats por
# categoría y filtrado por silenciadas.
# ────────────────────────────────────────────────────────────────────

@bp.route('/api/alertas/all', methods=['GET'])
def alertas_all():
    """Endpoint consolidado para tab Alertas. Devuelve 6 categorías:
      - mps_sin_stock · MPs activas con stock <= 0
      - mps_bajo_minimo · stock < stock_minimo (excluye sin_stock)
      - lotes_vencidos · fecha_venc < hoy (con stock > 0)
      - lotes_proximos · vence en próximos 30d
      - mees_bajo_minimo · MEE bajo mínimo
      - lotes_cuarentena · estado_lote IN (CUARENTENA, CUARENTENA_EXTENDIDA)

    Filtra alertas silenciadas activas. Devuelve stats + items por categoría.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()

    # Cargar set de (tipo, codigo) silenciados activos
    silenciados = set()
    try:
        for r in c.execute(
            """SELECT tipo_alerta, codigo_referencia FROM alertas_silenciadas
               WHERE activo = 1
                 AND (expira_at_utc IS NULL OR expira_at_utc > datetime('now','utc'))""",
        ).fetchall():
            silenciados.add((r[0], r[1]))
    except Exception:
        pass

    def _no_silenciado(tipo, cod):
        return (tipo, cod) not in silenciados

    # 1. MPs sin stock + bajo mínimo (1 query agrupada)
    rows_mp = c.execute(
        """SELECT mp.codigo_mp, COALESCE(mp.nombre_comercial,'') as nom,
                  COALESCE(mp.nombre_inci,'') as inci,
                  COALESCE(mp.proveedor,'') as prov,
                  COALESCE(mp.tipo,'') as subt,
                  COALESCE(mp.stock_minimo, 0) as smin,
                  COALESCE(s.stock, 0) as stock
           FROM maestro_mps mp
           LEFT JOIN (
               SELECT material_id,
                      SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                               WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                               ELSE 0 END) as stock
               FROM movimientos
               -- FIX 30-may-2026 · audit Planta/mínimos · agregar VENCIDO y
               -- AGOTADO a la exclusión (antes solo CUARENTENA/RECHAZADO/ANULADO)
               -- · un MP cuyo stock está vencido NO es usable → debe disparar
               -- "bajo mínimo". Unifica con /api/alertas-reabastecimiento.
               WHERE UPPER(COALESCE(estado_lote,'')) NOT IN
                     ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','ANULADO','VENCIDO','AGOTADO')
               GROUP BY material_id
           ) s ON mp.codigo_mp = s.material_id
           WHERE COALESCE(mp.activo, 1) = 1
             AND UPPER(COALESCE(mp.tipo_material,'MP')) = 'MP'
             AND COALESCE(mp.stock_minimo, 0) > 0""",
    ).fetchall()
    mps_sin_stock = []
    mps_bajo_minimo = []
    for r in rows_mp:
        cod, nom, inci, prov, subt, smin, stock = r
        smin = float(smin or 0); stock = float(stock or 0)
        if not _no_silenciado('mps_sin_stock' if stock <= 0 else 'mps_bajo_minimo', cod):
            continue
        item = {
            'codigo_mp': cod, 'nombre': nom or inci or cod,
            'nombre_inci': inci, 'proveedor': prov, 'subtipo': subt,
            'stock_minimo_g': smin, 'stock_actual_g': max(stock, 0),
            'deficit_g': max(smin - stock, 0),
            'cobertura_pct': round((stock / smin * 100) if smin > 0 else 0, 1),
        }
        if stock <= 0:
            mps_sin_stock.append(item)
        elif stock < smin:
            mps_bajo_minimo.append(item)
    mps_sin_stock.sort(key=lambda x: x['nombre'])
    mps_bajo_minimo.sort(key=lambda x: x['cobertura_pct'])

    # 2. Lotes vencidos + próximos
    rows_v = c.execute(
        """SELECT material_id, lote, MAX(material_nombre) as nombre,
                  MAX(fecha_vencimiento) as venc, MAX(proveedor) as prov,
                  SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                          WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                          ELSE 0 END) as stock
           FROM movimientos
           WHERE COALESCE(lote,'') != ''
             AND fecha_vencimiento IS NOT NULL AND fecha_vencimiento != ''
             AND UPPER(COALESCE(estado_lote,'')) NOT IN ('ANULADO','RECHAZADO')
           GROUP BY material_id, lote
           HAVING stock > 0""",
    ).fetchall()
    from datetime import date as _date
    hoy = _date.today()
    lotes_vencidos = []
    lotes_proximos = []
    for r in rows_v:
        mid, lote, nombre, venc, prov, stock = r
        if not venc:
            continue
        try:
            from datetime import datetime as _dt
            venc_d = _dt.strptime(str(venc)[:10], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue
        dias = (venc_d - hoy).days
        key_silen = f'{mid}::{lote}'
        if not _no_silenciado('lote_venc', key_silen):
            continue
        item = {
            'material_id': mid or '', 'lote': lote,
            'nombre': nombre or '', 'proveedor': prov or '',
            'cantidad_g': float(stock or 0),
            'fecha_vencimiento': str(venc)[:10],
            'dias_para_vencer': dias,
        }
        if dias < 0:
            lotes_vencidos.append(item)
        elif dias <= 30:
            lotes_proximos.append(item)
    lotes_vencidos.sort(key=lambda x: x['dias_para_vencer'])
    lotes_proximos.sort(key=lambda x: x['dias_para_vencer'])

    # 3. MEE bajo mínimo
    mees_bajo = []
    try:
        # stock CANÓNICO desde SUM(movimientos_mee) (no el cache, que driftea) ·
        # la alerta de quiebre debe usar el stock real (M5: display = decisión).
        rows_mee = c.execute(
            """SELECT m.codigo, m.descripcion, COALESCE(m.categoria,''),
                      COALESCE(m.proveedor,''),
                      COALESCE(m.stock_minimo, 0) as smin,
                      COALESCE((SELECT SUM(CASE WHEN mm.tipo='Entrada' THEN mm.cantidad
                                                WHEN mm.tipo='Salida'  THEN -mm.cantidad
                                                WHEN mm.tipo='Ajuste'  THEN mm.cantidad
                                                ELSE 0 END)
                                FROM movimientos_mee mm
                                WHERE mm.mee_codigo=m.codigo AND COALESCE(mm.anulado,0)=0), m.stock_actual, 0) as stock
               FROM maestro_mee m
               WHERE COALESCE(m.estado, 'Activo') = 'Activo'
                 AND COALESCE(m.stock_minimo, 0) > 0
               ORDER BY m.descripcion ASC""",
        ).fetchall()
        for r in rows_mee:
            cod, desc, cat, prov, smin, stock = r
            smin = float(smin or 0); stock = max(float(stock or 0), 0)
            if stock >= smin:   # solo los que están BAJO el mínimo (filtro en Python · stock canónico)
                continue
            if not _no_silenciado('mee_bajo_minimo', cod):
                continue
            mees_bajo.append({
                'codigo': cod, 'descripcion': desc or '',
                'categoria': cat, 'proveedor': prov,
                'stock_minimo': smin, 'stock_actual': stock,
                'deficit': max(smin - stock, 0),
                'cobertura_pct': round((stock / smin * 100) if smin > 0 else 0, 1),
            })
    except Exception:
        pass

    # 4. Lotes en cuarentena (case-insensitive)
    lotes_cuar = []
    try:
        rows_cuar = c.execute(
            """SELECT material_id, lote, MAX(material_nombre), MAX(proveedor),
                      MAX(fecha) as ult_fecha, MAX(numero_oc),
                      SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                              WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                              ELSE 0 END) as stock
               FROM movimientos
               WHERE UPPER(COALESCE(estado_lote,'')) IN ('CUARENTENA','CUARENTENA_EXTENDIDA')
               GROUP BY material_id, lote
               HAVING stock > 0
               ORDER BY ult_fecha ASC""",
        ).fetchall()
        for r in rows_cuar:
            mid, lote, nombre, prov, ult, oc, stock = r
            key = f'{mid}::{lote}'
            if not _no_silenciado('lote_cuarentena', key):
                continue
            lotes_cuar.append({
                'material_id': mid or '', 'lote': lote,
                'nombre': nombre or '', 'proveedor': prov or '',
                'fecha_ingreso': (ult or '')[:16].replace('T', ' '),
                'numero_oc': oc or '',
                'cantidad_g': float(stock or 0),
            })
    except Exception:
        pass

    # Agrupado por proveedor (para SOL combinada · solo bajo_minimo + sin_stock)
    by_proveedor = {}
    for it in mps_sin_stock + mps_bajo_minimo:
        prov = it['proveedor'] or '(sin proveedor)'
        by_proveedor.setdefault(prov, {
            'proveedor': prov,
            'items': [],
            'deficit_total_g': 0,
        })
        by_proveedor[prov]['items'].append({
            'codigo_mp': it['codigo_mp'],
            'nombre': it['nombre'],
            'deficit_g': it['deficit_g'],
            'urgencia': 'CRITICO' if it['cobertura_pct'] < 25 else (
                'URGENTE' if it['cobertura_pct'] < 50 else 'BAJO'),
        })
        by_proveedor[prov]['deficit_total_g'] += it['deficit_g']
    proveedores = sorted(by_proveedor.values(),
                         key=lambda g: -len(g['items']))

    return jsonify({
        'stats': {
            'mps_sin_stock': len(mps_sin_stock),
            'mps_bajo_minimo': len(mps_bajo_minimo),
            'lotes_vencidos': len(lotes_vencidos),
            'lotes_proximos': len(lotes_proximos),
            'mees_bajo_minimo': len(mees_bajo),
            'lotes_cuarentena': len(lotes_cuar),
            'total': (len(mps_sin_stock) + len(mps_bajo_minimo) +
                      len(lotes_vencidos) + len(lotes_proximos) +
                      len(mees_bajo) + len(lotes_cuar)),
            'silenciadas_activas': len(silenciados),
        },
        'mps_sin_stock': mps_sin_stock,
        'mps_bajo_minimo': mps_bajo_minimo,
        'lotes_vencidos': lotes_vencidos,
        'lotes_proximos': lotes_proximos,
        'mees_bajo_minimo': mees_bajo,
        'lotes_cuarentena': lotes_cuar,
        'agrupado_por_proveedor': proveedores,
    })


@bp.route('/api/alertas/silenciar', methods=['POST'])
def alertas_silenciar():
    """Sprint Alertas PRO · 20-may-2026 · silencia alerta con motivo.

    Body: {tipo_alerta, codigo_referencia, motivo, expira_dias?}
    Tipos válidos: mps_sin_stock, mps_bajo_minimo, lote_venc, lote_cuarentena,
                    mee_bajo_minimo
    """
    u, err, code = _require_session()
    if err:
        return err, code
    body = request.get_json(silent=True) or {}
    tipo = (body.get('tipo_alerta') or '').strip()
    cod = (body.get('codigo_referencia') or '').strip()
    motivo = (body.get('motivo') or '').strip()
    TIPOS_VALIDOS = ('mps_sin_stock', 'mps_bajo_minimo', 'lote_venc',
                      'lote_cuarentena', 'mee_bajo_minimo')
    if tipo not in TIPOS_VALIDOS:
        return jsonify({'error': f'tipo_alerta inválido · usar {TIPOS_VALIDOS}'}), 400
    if not cod:
        return jsonify({'error': 'codigo_referencia requerido'}), 400
    if len(motivo) < 10:
        return jsonify({'error': 'motivo requerido (≥10 chars)'}), 400
    expira_dias = body.get('expira_dias')
    expira_sql = "NULL"
    params_expira = []
    if expira_dias:
        try:
            d = int(expira_dias)
            if d > 0:
                expira_sql = "datetime('now','utc','+' || ? || ' days')"
                params_expira.append(str(d))
        except (ValueError, TypeError):
            pass
    conn = get_db(); c = conn.cursor()
    c.execute(
        f"""INSERT INTO alertas_silenciadas
              (tipo_alerta, codigo_referencia, motivo, silenciado_por,
               expira_at_utc, activo)
            VALUES (?, ?, ?, ?, {expira_sql}, 1)""",
        [tipo, cod, motivo[:500], u] + params_expira,
    )
    new_id = c.lastrowid
    try:
        audit_log(c, usuario=u, accion='SILENCIAR_ALERTA',
                  tabla='alertas_silenciadas', registro_id=new_id,
                  despues={'tipo': tipo, 'codigo': cod,
                           'motivo': motivo[:200],
                           'expira_dias': expira_dias})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'id': new_id,
                    'mensaje': f'Alerta silenciada · {cod}'})


@bp.route('/api/alertas/silenciar/<int:silen_id>', methods=['DELETE'])
def alertas_desilenciar(silen_id):
    """Re-activar alerta silenciada (activo=0)."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    row = c.execute(
        "SELECT tipo_alerta, codigo_referencia FROM alertas_silenciadas WHERE id = ?",
        (silen_id,)).fetchone()
    if not row:
        return jsonify({'error': 'no existe'}), 404
    c.execute("UPDATE alertas_silenciadas SET activo = 0 WHERE id = ?", (silen_id,))
    try:
        audit_log(c, usuario=u, accion='DESILENCIAR_ALERTA',
                  tabla='alertas_silenciadas', registro_id=silen_id,
                  despues={'tipo': row[0], 'codigo': row[1]})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True})


@bp.route('/api/alertas-reabastecimiento')
def alertas_reabastecimiento():
    conn = get_db(); c = conn.cursor()
    # MPs bajo minimo
    # Subconsulta: las columnas de maestro_mps y los alias no se pueden
    # usar en WHERE/ORDER BY de la query agrupada en Postgres · envueltas
    # en un FROM (...) se vuelven columnas reales (portátil SQLite/PG).
    # INVIMA-FIX · 21-may-2026 · alertas excluyen CUARENTENA/RECHAZADO
    # Antes: MP con 1000g en CUARENTENA contaba como stock · alerta NO se
    # disparaba pero FEFO bloqueaba producción · inconsistencia grave.
    c.execute("""SELECT material_id, nombre, proveedor, stock_minimo,
                        stock_actual, tipo_material, subtipo FROM (
                   SELECT m.material_id AS material_id,
                          COALESCE(MAX(mp.nombre_comercial), MAX(m.material_nombre)) AS nombre,
                          COALESCE(MAX(mp.proveedor),'') AS proveedor,
                          COALESCE(MAX(mp.stock_minimo),0) AS stock_minimo,
                          SUM(CASE
                                WHEN m.tipo IN ('Entrada','Ajuste +','Ajuste') THEN m.cantidad
                                WHEN m.tipo IN ('Salida','Ajuste -') THEN -m.cantidad
                                ELSE 0
                              END) AS stock_actual,
                          'MP' AS tipo_material,
                          COALESCE(MAX(mp.tipo),'') AS subtipo
                   FROM movimientos m
                   LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
                   WHERE m.estado_lote IS NULL
                      OR UPPER(COALESCE(m.estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO')
                   GROUP BY m.material_id
                 ) sub
                 WHERE stock_actual < stock_minimo AND stock_minimo > 0
                 ORDER BY (stock_actual/stock_minimo) ASC""")
    rows_mp = c.fetchall()
    # MEE bajo minimo
    c.execute("""SELECT codigo, descripcion, '' as proveedor, stock_minimo,
                        stock_actual, 'MEE' as tipo_material, categoria as subtipo
                 FROM maestro_mee
                 WHERE estado='Activo' AND stock_minimo > 0 AND stock_actual < stock_minimo
                 ORDER BY (stock_actual/stock_minimo) ASC""")
    rows_mee = c.fetchall()
    # ABASTECIMIENTO-FIX · 22-may-2026 · dedup cola pendiente (#8 audit 22-may)
    # · Antes: alertas crónicas para MPs ya en cola · ruido en dashboard
    # · Ahora: muestra 'en_cola_g' · deficit considera cola · alerta solo si
    #   stock+cola aún no cubre mínimo
    # PERF-FIX 23-may-2026 · auditoría · antes _pendiente_en_compras_g se
    # llamaba por cada alerta (50 alertas × 2 SELECTs = 100 round-trips PG) ·
    # ahora precomputamos un dict {codigo_mp: pendiente_g} con 2 queries
    # agregadas (usa indexes idx_sci_codigo_mp + idx_oci_codigo_mp · mig 152)
    pendientes_dict = {}
    try:
        for cm, g_pend in c.execute(
            """SELECT sci.codigo_mp, COALESCE(SUM(sci.cantidad_g), 0)
               FROM solicitudes_compra_items sci
               JOIN solicitudes_compra sc ON sc.numero = sci.numero
               WHERE sc.estado IN ('Pendiente','Aprobada')
                 AND COALESCE(sc.numero_oc,'') = ''
                 AND sci.codigo_mp IS NOT NULL AND TRIM(sci.codigo_mp) != ''
               GROUP BY sci.codigo_mp"""
        ).fetchall():
            pendientes_dict[str(cm).strip()] = float(g_pend or 0)
    except Exception:
        pass
    try:
        for cm, g_pend in c.execute(
            """SELECT oci.codigo_mp,
                      COALESCE(SUM(oci.cantidad_g - COALESCE(oci.cantidad_recibida_g,0)), 0)
               FROM ordenes_compra_items oci
               JOIN ordenes_compra oc ON oc.numero_oc = oci.numero_oc
               WHERE oc.estado IN ('Borrador','Revisada','Autorizada','Parcial')
                 AND oci.codigo_mp IS NOT NULL AND TRIM(oci.codigo_mp) != ''
               GROUP BY oci.codigo_mp"""
        ).fetchall():
            k = str(cm).strip()
            pendientes_dict[k] = pendientes_dict.get(k, 0.0) + float(g_pend or 0)
    except Exception:
        pass
    alertas = []
    for r in list(rows_mp) + list(rows_mee):
        stock_actual = round(r[4] or 0, 1)
        stock_minimo = round(r[3], 1)
        cod = r[0] or ''
        en_cola_g = round(pendientes_dict.get(str(cod).strip(), 0.0), 1) if cod else 0
        deficit_neto = round(max(stock_minimo - stock_actual - en_cola_g, 0), 1)
        alertas.append({'codigo_mp': cod, 'nombre': r[1] or '', 'proveedor': r[2] or '',
                        'stock_minimo': stock_minimo, 'stock_actual': max(stock_actual, 0),
                        'en_cola_g': en_cola_g,
                        'deficit': deficit_neto,
                        'tipo': r[5] or 'MP', 'subtipo': r[6] or '',
                        'cubierto_por_cola': en_cola_g > 0 and deficit_neto <= 0})
    alertas.sort(key=lambda x: x['stock_actual']/x['stock_minimo'] if x['stock_minimo'] else 1)
    return jsonify({'alertas': alertas, 'total': len(alertas)})

@bp.route('/api/stock')
def get_stock():
    # SEC-FIX · 21-may-2026 · auth obligatorio (antes endpoint público)
    u, err, code = _require_session()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    # ABASTECIMIENTO-FIX · 22-may-2026 · Ajustes + Ajuste+ suman · Salida + Ajuste- restan
    # B2 (12-jun): este es el FISICO TOTAL a proposito (incluye cuarentena/retenido).
    # NO es "disponible para producir": para disponibilidad usar /api/lotes (excluye
    # estados no-producibles · A1) o el resolver canonico. No usar este KPI para decidir
    # compras/produccion sin restar lo retenido.
    c.execute("""SELECT material_id, material_nombre,
                   SUM(CASE WHEN tipo IN ('Entrada','Ajuste +','Ajuste') THEN cantidad ELSE 0 END) as entradas,
                   SUM(CASE WHEN tipo IN ('Salida','Ajuste -') THEN cantidad ELSE 0 END) as salidas,
                   SUM(CASE
                         WHEN tipo IN ('Entrada','Ajuste +','Ajuste') THEN cantidad
                         WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad
                         ELSE 0
                       END) as stock_actual
                 FROM movimientos GROUP BY material_id, material_nombre ORDER BY material_nombre""")
    rows = c.fetchall()
    return jsonify({'items': [{'material_id':r[0],'material_nombre':r[1],'entradas':round(r[2] or 0,2),'salidas':round(r[3] or 0,2),'stock_actual':round(r[4] or 0,2)} for r in rows]})

@bp.route('/api/lotes')
def get_lotes():
    # SEC-FIX · 21-may-2026 · auth obligatorio (antes endpoint público)
    _u, _err, _code = _require_session()
    if _err: return _err, _code
    """Lista lotes de Materia Prima con stock disponible (stock_neto > 0).

    Sebastian 5-may-2026 (audit zero-error Bodega MP): 3 fixes criticos.

    Bug 1 — MEEs aparecian en lista MP:
      ANTES no filtraba por tipo_material='MP'. Si por error operativo un
      envase/tapa/etiqueta se registró en maestro_mps + movimientos (en
      lugar de maestro_mee + movimientos_mee), aparecia en Bodega MP.
      Fix: WHERE COALESCE(mp.tipo_material,'MP')='MP'. El COALESCE permite
      que MPs legacy sin tipo_material set sigan apareciendo (default MP).

    Bug 2 — lotes consumidos visibles:
      ANTES HAVING stock_neto > -999999 mostraba lotes con stock 0 o
      negativo (drift). Catalina veia lotes ya consumidos como
      'disponibles' y planificaba sobre stock fantasma.
      Fix: HAVING stock_neto > 0 — solo lotes con stock real disponible.

    Bug 3 — sin paginacion (cargaba 5k+ lotes en mobile):
      ANTES sin LIMIT · query devolvia toda la tabla agrupada · 500ms+
      en mobile.
      Fix: paginacion OPCIONAL via query params · backwards-compatible.
        ?limit=N (default sin limite si no se especifica)
        ?offset=N (default 0)
        ?solo_criticos=1 (solo lotes a vencer en 30d o vencidos)
      Si no llegan params → comportamiento legacy (todos).
    """
    from datetime import date; hoy = date.today().isoformat()

    # Paginacion opcional · backwards-compatible
    try:
        limit = int(request.args.get('limit') or 0)
    except (ValueError, TypeError):
        limit = 0
    try:
        offset = int(request.args.get('offset') or 0)
    except (ValueError, TypeError):
        offset = 0
    if limit < 0: limit = 0
    if offset < 0: offset = 0
    if limit > 5000: limit = 5000
    solo_criticos = (request.args.get('solo_criticos') or '').strip().lower() in ('1','true','yes')
    # Sebastián 3-jun-2026: la Bodega MP debe listar TODAS las materias primas
    # (tengan o no inventario) con su stock mínimo. Opt-in para no cambiar el
    # comportamiento por defecto del endpoint (otros consumidores + tests).
    incluir_sin_stock = (request.args.get('incluir_sin_stock') or '').strip().lower() in ('1','true','yes')

    conn = get_db(); c = conn.cursor()

    # Construir query con filtros opcionales
    sql = """SELECT m.material_id, MAX(m.material_nombre) as material_nombre, m.lote,
                        SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as stock_neto,
                        MAX(m.fecha_vencimiento) as fecha_vencimiento,
                        MAX(m.estanteria) as estanteria, MAX(m.posicion) as posicion,
                        MAX(m.proveedor) as proveedor, MAX(m.estado_lote) as estado_lote,
                        COALESCE(MAX(mp.nombre_inci),'') as inci,
                        COALESCE(MAX(mp.tipo),'') as tipo,
                        COALESCE(MAX(mp.stock_minimo),0) as smin
                 FROM movimientos m LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
                 WHERE UPPER(COALESCE(mp.tipo_material,'MP'))='MP'
                   AND (m.estado_lote IS NULL OR UPPER(COALESCE(m.estado_lote,'')) NOT IN
                        ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO'))
                 GROUP BY m.material_id, m.lote
                 HAVING stock_neto > 0.01"""
    # A1 (Sebastian 12-jun · decision: excluir): la Bodega MP muestra solo stock
    # USABLE · antes listaba lotes en cuarentena/rechazado/vencido/agotado como
    # disponibles e inflaba el total -> enmascaraba quiebres (M5). Ahora alineado
    # con conteo_ciclico + FEFO + descuento. El stock retenido vive en su vista
    # de Cuarentena/Calidad (/api/lotes/cuarentena).
    # Sebastian 12-jun: umbral 0.01g (no >0). Un lote ya gastado deja un residuo
    # flotante diminuto (0.004g del redondeo del descuento) que pasaba >0 pero se
    # mostraba como "0" y nunca desaparecia. <=0.01g = consumido -> sale de la vista
    # y de FEFO (siempre se elige donde HAY · al agotarse, el lote desaparece solo).
    if solo_criticos:
        # Solo lotes vencidos o que vencen en proximos 30d (con fecha_venc set)
        sql += (" AND MAX(m.fecha_vencimiento) IS NOT NULL"
                " AND MAX(m.fecha_vencimiento) != ''"
                " AND MAX(m.fecha_vencimiento) <= date('now', '-5 hours', '+30 day')")
    sql += " ORDER BY material_nombre ASC, fecha_vencimiento ASC"
    if limit > 0:
        sql += f" LIMIT {limit} OFFSET {offset}"
    c.execute(sql)
    rows = c.fetchall()

    # Sebastian 5-may-2026 (Luis Enrique reportó "Solicitar no hace nada"):
    # cargar set de codigo_mp con solicitudes Pendientes para que UI muestre
    # badge "Solicitada" en cada lote afectado · feedback visible en Planta
    # de que el boton tuvo efecto.
    codigos_con_solicitud = set()
    try:
        for sr in c.execute("""
            SELECT DISTINCT sci.codigo_mp
            FROM solicitudes_compra_items sci
            JOIN solicitudes_compra sc ON sc.numero = sci.numero
            WHERE sc.estado='Pendiente'
              AND sci.codigo_mp IS NOT NULL
              AND sci.codigo_mp != ''
        """).fetchall():
            codigos_con_solicitud.add(sr[0])
    except Exception as _e:
        __import__('logging').getLogger('inventario').warning(
            "get_lotes solicitudes_pendientes lookup falló: %s", _e,
        )

    # Sebastián 9-may-2026: stock_minimo es a nivel del MP completo (suma
    # de lotes), NO de cada lote individual. Antes la UI pintaba como
    # "bajo mínimo" un lote con menos cantidad que el mínimo · falso
    # positivo si el MP tiene OTROS lotes con stock suficiente. Ahora
    # calculamos el TOTAL por material y lo devolvemos en cada fila para
    # que el frontend compare contra ese total.
    totales_mp = {}
    for r in rows:
        mid_r = r[0] or ''
        totales_mp[mid_r] = totales_mp.get(mid_r, 0) + (r[3] or 0)

    result = []
    for r in rows:
        mid,mnm,lote,cant,fvenc,est,pos,prov,estado,inci,tipo,smin = r
        dias,alerta = None,'ok'
        if fvenc and len(str(fvenc))>=10:
            try:
                from datetime import datetime as dt2
                dias=(dt2.strptime(str(fvenc)[:10],'%Y-%m-%d').date()-dt2.strptime(hoy,'%Y-%m-%d').date()).days
                alerta='vencido' if dias<0 else ('critico' if dias<=30 else ('proximo' if dias<=90 else 'ok'))
            except (ValueError, TypeError):
                # Fecha mal formateada — dejar como 'ok' silencio aceptable
                # (no es crítico para el flujo de stock).
                pass
        result.append({'material_id':mid or '','nombre_inci':inci,'material_nombre':mnm or '',
                       'tipo':tipo,'proveedor':prov or '','stock_min_g':round(smin,1),
                       'stock_total_mp_g':round(totales_mp.get(mid or '', 0), 2),
                       'lote':lote or '','cantidad_g':round(cant or 0,2),'cantidad_kg':round((cant or 0)/1000,3),
                       'estanteria':est or '','posicion':pos or '',
                       'fecha_vencimiento':str(fvenc)[:10] if fvenc else '',
                       'dias_para_vencer':dias,'estado_lote':estado or '','alerta':alerta,
                       'tiene_solicitud_pendiente': (mid or '') in codigos_con_solicitud})

    # Sebastián 3-jun-2026 · incluir TODAS las MPs del maestro (con o sin stock)
    # con su stock mínimo. Las que NO tienen lote con stock se agregan como fila
    # "sin stock" (cantidad 0). Opt-in (incluir_sin_stock) · no aplica a las
    # vistas filtradas (solo_criticos). Respeta tipo_material='MP' y activo=1.
    if incluir_sin_stock and not solo_criticos:
        try:
            _ya = {x['material_id'] for x in result}
            _cat = c.execute(
                """SELECT codigo_mp,
                          COALESCE(NULLIF(TRIM(nombre_comercial),''),
                                   NULLIF(TRIM(nombre_inci),''), codigo_mp) AS nombre,
                          COALESCE(nombre_inci,'') AS inci,
                          COALESCE(tipo,'') AS tipo,
                          COALESCE(stock_minimo,0) AS smin
                     FROM maestro_mps
                    WHERE UPPER(COALESCE(tipo_material,'MP'))='MP'
                      AND COALESCE(activo,1)=1""").fetchall()
            for cr in _cat:
                _cod = cr[0]
                if not _cod or _cod in _ya:
                    continue
                result.append({
                    'material_id': _cod, 'nombre_inci': cr[2], 'material_nombre': cr[1],
                    'tipo': cr[3], 'proveedor': '', 'stock_min_g': round(cr[4] or 0, 1),
                    'stock_total_mp_g': 0, 'lote': '', 'cantidad_g': 0, 'cantidad_kg': 0,
                    'estanteria': '', 'posicion': '', 'fecha_vencimiento': '',
                    'dias_para_vencer': None, 'estado_lote': '', 'alerta': 'sin_stock',
                    'tiene_solicitud_pendiente': _cod in codigos_con_solicitud})
            result.sort(key=lambda x: (x.get('material_nombre') or '').lower())
        except Exception as _e:
            __import__('logging').getLogger('inventario').warning(
                "get_lotes incluir_sin_stock fallo: %s", _e)

    # Si llegó paginacion, incluir conteo total para que UI sepa cuantas
    # paginas hay. Si no, total = len(result) (legacy compat).
    response = {'lotes': result, 'total': len(result)}
    if limit > 0:
        try:
            count_sql = """
                SELECT COUNT(*) FROM (
                  SELECT m.material_id, m.lote,
                         SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as sn
                  FROM movimientos m LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
                  WHERE UPPER(COALESCE(mp.tipo_material,'MP'))='MP'
                    AND (m.estado_lote IS NULL OR UPPER(COALESCE(m.estado_lote,'')) NOT IN
                         ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO'))
                  GROUP BY m.material_id, m.lote
                  HAVING sn > 0.01
                )
            """
            total_count = c.execute(count_sql).fetchone()[0]
            response['total_disponibles'] = total_count
            response['offset'] = offset
            response['limit'] = limit
            response['has_more'] = (offset + len(result)) < total_count
        except Exception as _e:
            __import__('logging').getLogger('inventario').warning(
                "get_lotes count fallo: %s", _e,
            )
    return jsonify(response)

# ────────────────────────────────────────────────────────────────────
# Maestro MPs · Detector + Unificador de duplicados · Sebastián
# 20-may-2026. Caso real: 'purisil' tiene 2 codigos distintos. Normaliza
# nombre + INCI, agrupa, permite unificar con audit_log completo.
# ────────────────────────────────────────────────────────────────────

def _normalizar_nombre_mp(s):
    """Normaliza nombre MP para detección de duplicados.
    - lowercase, strip
    - quitar tildes
    - quitar puntos, comas, paréntesis, guiones, slashes
    - colapsar espacios múltiples
    """
    import re, unicodedata
    if not s:
        return ''
    n = str(s).lower().strip()
    n = unicodedata.normalize('NFD', n)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    n = re.sub(r'[.,;:()\[\]/\\\-_*]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


@bp.route('/api/maestro-mps/duplicados-deteccion', methods=['GET'])
def maestro_mps_duplicados_deteccion():
    """Detecta MPs duplicadas (nombre_comercial o nombre_inci similares
    con codigos distintos). Solo lectura · admin.

    Devuelve grupos con ≥ 2 codigos que normalizan al mismo nombre. Para
    cada variante incluye: codigo_mp, nombre, INCI, proveedor, activo,
    stock_actual_g (suma de movimientos), n_movimientos, n_lotes,
    n_formulas (usado en formula_items), n_sols (en SOLs).

    Usa esto para identificar duplicados antes de unificar con
    /api/maestro-mps/unificar.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'solo admin'}), 403

    conn = get_db(); c = conn.cursor()
    mps = c.execute(
        """SELECT codigo_mp, nombre_comercial, nombre_inci, proveedor,
                  COALESCE(activo,1) as activo
           FROM maestro_mps
           ORDER BY nombre_comercial ASC""",
    ).fetchall()
    grupos = {}
    for r in mps:
        cod, nom, inci, prov, activo = r
        key_nom = _normalizar_nombre_mp(nom)
        key_inci = _normalizar_nombre_mp(inci)
        # Clave compuesta: si ambos coinciden cuentan como mismo grupo.
        # Si solo coincide nombre o solo INCI, también es grupo
        # candidato (puede ser ambiguo, el usuario decidirá).
        keys = []
        if key_nom: keys.append(('nombre', key_nom))
        if key_inci and key_inci != key_nom: keys.append(('inci', key_inci))
        for ktipo, k in keys:
            grupos.setdefault((ktipo, k), []).append(
                {'codigo_mp': cod, 'nombre_comercial': nom or '',
                 'nombre_inci': inci or '', 'proveedor': prov or '',
                 'activo': bool(activo)})

    # Filtrar solo grupos con > 1 codigo distinto
    duplicados = []
    for (ktipo, k), variantes in grupos.items():
        codigos = {v['codigo_mp'] for v in variantes}
        if len(codigos) < 2:
            continue
        # Para cada variante, traer stats: stock, movs, lotes, formulas, sols
        for v in variantes:
            stock_row = c.execute(
                # M1 (12-jun): CASE canonico · antes 'tipo=Entrada ELSE -cantidad'
                # restaba los 'Ajuste'/'Ajuste +'/'entrada' minuscula (tipos reales)
                # -> stock drift en esta vista de duplicados (regla #4 / M16).
                "SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad "
                "WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END),0), "
                "COUNT(*), COUNT(DISTINCT COALESCE(lote,'')) "
                "FROM movimientos WHERE material_id = ?",
                (v['codigo_mp'],),
            ).fetchone()
            v['stock_actual_g'] = round(float(stock_row[0] or 0), 2)
            v['n_movimientos'] = int(stock_row[1] or 0)
            v['n_lotes'] = int(stock_row[2] or 0)
            try:
                n_form = c.execute(
                    "SELECT COUNT(*) FROM formula_items WHERE material_id = ?",
                    (v['codigo_mp'],)).fetchone()[0]
            except Exception:
                n_form = 0
            v['n_formulas'] = int(n_form or 0)
            try:
                n_sol = c.execute(
                    "SELECT COUNT(*) FROM solicitudes_compra_items WHERE codigo_mp = ?",
                    (v['codigo_mp'],)).fetchone()[0]
            except Exception:
                n_sol = 0
            v['n_sols'] = int(n_sol or 0)
        duplicados.append({
            'tipo_match': ktipo,
            'nombre_normalizado': k,
            'variantes': sorted(variantes,
                                key=lambda x: (-x['n_movimientos'],
                                               -x['stock_actual_g'])),
        })

    # Ordenar grupos por cantidad de impacto
    duplicados.sort(
        key=lambda g: -sum(v['n_movimientos'] for v in g['variantes']))
    return jsonify({
        'grupos': duplicados,
        'total_grupos': len(duplicados),
        'total_mps_afectadas': sum(len(g['variantes']) for g in duplicados),
    })


# Tablas que referencian material_id / codigo_mp · USADAS por el
# unificador para reescribir referencias del codigo_origen al canonico.
# Tupla: (tabla, columna).
_TABLAS_REF_MP = [
    ('movimientos', 'material_id'),
    ('formula_items', 'material_id'),
    ('solicitudes_compra_items', 'codigo_mp'),
    ('ordenes_compra_items', 'material_id'),
    ('mp_lead_time_config', 'material_id'),
    # Audit 3-jun · mp_formula_bridge NO tiene columna material_id (eran
    # formula_material_id / bodega_material_id) → la unify nunca repuntaba el
    # bridge y dejaba filas colgadas que resolvían a stock 0. Se corrigen ambas.
    ('mp_formula_bridge', 'formula_material_id'),
    ('mp_formula_bridge', 'bodega_material_id'),
    ('precios_mp_historico', 'material_id'),
    ('conteo_items', 'material_id'),
    ('conteo_ciclico_calendario', 'material_id'),
    ('conteo_ciclico_config', 'material_id'),
    ('ebr_pesajes', 'material_id'),
    ('especificaciones_mp', 'material_id'),
    ('alertas', 'codigo_mp'),
]


@bp.route('/api/maestro-mps/unificar', methods=['POST'])
def maestro_mps_unificar():
    """Unifica códigos duplicados de MP en un único canónico.

    Body: { canonico: str, codigos_a_unir: [str, ...],
            dry_run: bool (default true), token: str }

    Por defecto dry_run=true · cuenta cuántas filas se actualizarían en
    cada tabla sin tocar nada. Si dry_run=false, requiere token válido
    'UNIFICAR_MP_2026' y aplica TODO en transacción:

    1. UPDATE tabla.col = canonico WHERE col IN (codigos_a_unir) en cada
       una de las _TABLAS_REF_MP.
    2. UPDATE maestro_mps SET activo=0 WHERE codigo_mp IN (codigos_a_unir)
       AND codigo_mp != canonico · NO los borra (auditoría histórica).
    3. audit_log por cada tabla afectada + log master MERGE.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'solo admin'}), 403

    body = request.get_json(silent=True) or {}
    canonico = (body.get('canonico') or '').strip()
    a_unir_raw = body.get('codigos_a_unir') or []
    if not isinstance(a_unir_raw, list):
        return jsonify({'error': 'codigos_a_unir debe ser lista'}), 400
    a_unir = [str(x).strip() for x in a_unir_raw if str(x).strip()]
    if not canonico:
        return jsonify({'error': 'canonico requerido'}), 400
    if not a_unir:
        return jsonify({'error': 'codigos_a_unir vacía'}), 400
    if canonico in a_unir:
        a_unir = [x for x in a_unir if x != canonico]
    if not a_unir:
        return jsonify({'error': 'codigos_a_unir solo contiene al canonico'}), 400
    dry_run = bool(body.get('dry_run', True))
    token = (body.get('token') or '').strip()
    if not dry_run and token != 'UNIFICAR_MP_2026':
        return jsonify({
            'error': 'token inválido para apply',
            'hint': 'usa token=UNIFICAR_MP_2026 cuando dry_run=false',
        }), 403

    conn = get_db(); c = conn.cursor()
    # Validar canonico existe y está activo
    canon_row = c.execute(
        "SELECT codigo_mp, nombre_comercial, COALESCE(activo,1) "
        "FROM maestro_mps WHERE codigo_mp = ?", (canonico,)).fetchone()
    if not canon_row:
        return jsonify({'error': f'canonico {canonico} no existe en maestro_mps'}), 404
    # Validar que cada codigo_a_unir existe
    a_unir_existentes = []
    for cod in a_unir:
        r = c.execute(
            "SELECT codigo_mp FROM maestro_mps WHERE codigo_mp = ?",
            (cod,)).fetchone()
        if r:
            a_unir_existentes.append(cod)
    if not a_unir_existentes:
        return jsonify({'error': 'ningún codigo_a_unir existe en maestro_mps'}), 404

    placeholders = ','.join('?' * len(a_unir_existentes))
    plan = {}  # tabla → cantidad de filas que serían actualizadas

    if dry_run:
        # Contar (no UPDATE)
        for tabla, col in _TABLAS_REF_MP:
            try:
                n = c.execute(
                    f"SELECT COUNT(*) FROM {tabla} WHERE {col} IN ({placeholders})",
                    a_unir_existentes,
                ).fetchone()[0]
                plan[tabla] = int(n or 0)
            except Exception as _e:
                plan[tabla] = f'ERR: {str(_e)[:80]}'
        return jsonify({
            'dry_run': True,
            'canonico': canonico,
            'codigos_a_unir': a_unir_existentes,
            'plan_updates_por_tabla': plan,
            'total_filas_a_actualizar': sum(
                v for v in plan.values() if isinstance(v, int)),
            'mps_a_desactivar': len(a_unir_existentes),
        })

    # APPLY · transacción
    afectados = {}
    try:
        for tabla, col in _TABLAS_REF_MP:
            try:
                cur = c.execute(
                    f"UPDATE {tabla} SET {col} = ? WHERE {col} IN ({placeholders})",
                    [canonico] + a_unir_existentes,
                )
                afectados[tabla] = cur.rowcount or 0
            except Exception as _e:
                __import__('logging').getLogger('inventario').warning(
                    'unificar fallo tabla %s: %s', tabla, _e)
                afectados[tabla] = f'ERR: {str(_e)[:80]}'
        # Desactivar (NO borrar) los codigos viejos
        cur2 = c.execute(
            f"UPDATE maestro_mps SET activo = 0 "
            f"WHERE codigo_mp IN ({placeholders})",
            a_unir_existentes,
        )
        afectados['maestro_mps_desactivados'] = cur2.rowcount or 0
        # audit_log master
        from audit_helpers import audit_log
        audit_log(c, usuario=u, accion='UNIFICAR_MP_DUPLICADOS',
                  tabla='maestro_mps', registro_id=canonico,
                  despues={'canonico': canonico,
                           'codigos_unidos': a_unir_existentes,
                           'filas_actualizadas': afectados})
        conn.commit()
    except Exception as _e:
        try: conn.rollback()
        except Exception: pass
        return jsonify({
            'error': f'transacción falló: {_e}',
            'afectados_parcial': afectados,
        }), 500

    return jsonify({
        'ok': True,
        'canonico': canonico,
        'codigos_unidos': a_unir_existentes,
        'filas_actualizadas': afectados,
        'mensaje': f'{len(a_unir_existentes)} MPs unificadas en {canonico}',
    })


@bp.route('/api/movimientos/recientes', methods=['GET'])
def movimientos_recientes():
    """Sprint Movimientos PRO · 20-may-2026 · paginado + filtros server-side.

    Antes el frontend bajaba TODOS los movimientos via /api/movimientos
    y filtraba en JS · perf horrible con 50k+ filas.

    Query params:
      limit (default 50, max 500)
      offset (default 0)
      q: busca en material_nombre / material_id / lote / observaciones
      tipo: Entrada | Salida | Ajuste (case-insensitive)
      desde: YYYY-MM-DD (>=)
      hasta: YYYY-MM-DD (<=)
      solo_anulados: 1 para mostrar solo anulados
    """
    u, err, code = _require_session()
    if err:
        return err, code
    try:
        limit = max(1, min(int(request.args.get('limit', 50)), 500))
    except (ValueError, TypeError):
        limit = 50
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    q = (request.args.get('q') or '').strip().lower()
    tipo = (request.args.get('tipo') or '').strip()
    desde = (request.args.get('desde') or '').strip()
    hasta = (request.args.get('hasta') or '').strip()
    solo_anul = (request.args.get('solo_anulados') or '').strip() in ('1','true','yes')

    where = ['1=1']
    params = []
    if q:
        q_safe = q.replace('\\','\\\\').replace('%','\\%').replace('_','\\_')
        like = f'%{q_safe}%'
        where.append(
            "(LOWER(COALESCE(material_nombre,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(material_id,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(lote,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(observaciones,'')) LIKE ? ESCAPE '\\')",
        )
        params.extend([like, like, like, like])
    if tipo:
        where.append("UPPER(COALESCE(tipo,'')) = UPPER(?)")
        params.append(tipo)
    if desde:
        where.append("fecha >= ?"); params.append(desde)
    if hasta:
        where.append("fecha <= ?"); params.append(hasta + ' 23:59:59')
    if solo_anul:
        where.append("COALESCE(observaciones,'') LIKE '[ANULADO]%'")
    where_sql = ' AND '.join(where)
    conn = get_db(); c = conn.cursor()
    try:
        total = c.execute(
            f"SELECT COUNT(*) FROM movimientos WHERE {where_sql}", params,
        ).fetchone()[0]
    except Exception:
        total = 0
    try:
        rows = c.execute(
            f"""SELECT id, material_id, material_nombre, cantidad, tipo,
                       COALESCE(lote,'') as lote, fecha,
                       COALESCE(observaciones,'') as obs,
                       COALESCE(operador,'') as oper,
                       COALESCE(proveedor,'') as prov,
                       COALESCE(numero_oc,'') as numero_oc,
                       COALESCE(numero_factura,'') as numero_factura,
                       COALESCE(estado_lote,'') as estado_lote
                FROM movimientos
                WHERE {where_sql}
                ORDER BY fecha DESC, id DESC
                LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
    except Exception as _e:
        __import__('logging').getLogger('inventario').warning(
            'movimientos_recientes query fallo: %s', _e)
        rows = []
    items = [{
        'id': r[0], 'material_id': r[1] or '',
        'material_nombre': r[2] or '', 'cantidad': float(r[3] or 0),
        'tipo': r[4] or '', 'lote': r[5] or '', 'fecha': r[6] or '',
        'observaciones': r[7] or '', 'operador': r[8] or '',
        'proveedor': r[9] or '', 'numero_oc': r[10] or '',
        'numero_factura': r[11] or '', 'estado_lote': r[12] or '',
        'anulado': (r[7] or '').startswith('[ANULADO]'),
    } for r in rows]
    return jsonify({
        'items': items, 'total': total,
        'limit': limit, 'offset': offset,
        'has_more': (offset + len(items)) < total,
    })


@bp.route('/api/lotes/<material_id>/<path:lote>/movimientos', methods=['GET'])
def get_lote_movimientos(material_id, lote):
    """Sprint Bodega MP PRO · 20-may-2026 · Sebastián.

    Devuelve los movimientos de UN lote específico (material_id + lote).
    Antes la UI llamaba `/api/movimientos` que trae TODA la tabla y
    filtraba en JS · perf horrible con miles de filas.

    Acepta `_SIN_LOTE_` como marcador para movimientos sin lote.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db()
    c = conn.cursor()
    lote_real = '' if lote == '_SIN_LOTE_' else lote
    rows = c.execute(
        """SELECT id, material_id, material_nombre, cantidad, tipo, lote,
                  fecha, observaciones, operador, proveedor,
                  COALESCE(estado_lote,''), COALESCE(fecha_vencimiento,'')
           FROM movimientos
           WHERE material_id = ? AND COALESCE(lote,'') = ?
           ORDER BY fecha DESC, id DESC
           LIMIT 500""",
        (material_id, lote_real),
    ).fetchall()
    items = [{
        'id': r[0], 'material_id': r[1] or '',
        'material_nombre': r[2] or '',
        'cantidad': float(r[3] or 0), 'tipo': r[4] or '',
        'lote': r[5] or '', 'fecha': r[6] or '',
        'observaciones': r[7] or '', 'operador': r[8] or '',
        'proveedor': r[9] or '',
        'estado_lote': r[10] or '', 'fecha_vencimiento': str(r[11] or ''),
    } for r in rows]
    # resumen por estado_lote · para ver de un vistazo por qué producción excluye
    _por_estado = {}
    for it in items:
        e = it['estado_lote'] or '(sin estado)'
        d = _por_estado.setdefault(e, {'entradas': 0.0, 'salidas': 0.0})
        if it['tipo'] in ('Entrada', 'entrada', 'ENTRADA', 'Ajuste +', 'Ajuste'):
            d['entradas'] += it['cantidad']
        elif it['tipo'] in ('Salida', 'salida', 'SALIDA', 'Ajuste -'):
            d['salidas'] += it['cantidad']
    resumen = {e: {'neto_g': round(v['entradas'] - v['salidas'], 2)} for e, v in _por_estado.items()}
    return jsonify({'movimientos': items, 'total': len(items),
                    'material_id': material_id, 'lote': lote_real,
                    'resumen_por_estado': resumen})


@bp.route('/api/proveedores-unicos', methods=['GET'])
def proveedores_unicos():
    """Lista de proveedores únicos para autocompletado en edición de lote.

    Une los valores de movimientos.proveedor + maestro_mps.proveedor para
    sugerir solo nombres ya conocidos y evitar duplicados por typo
    ('Inchemical' vs 'INCHEMICAL'). Caso-sensitive para conservar el
    formato canónico que el usuario ya escogió.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db()
    c = conn.cursor()
    proveedores = set()
    try:
        for row in c.execute("SELECT DISTINCT proveedor FROM movimientos "
                             "WHERE proveedor IS NOT NULL AND proveedor != ''"):
            proveedores.add(row[0].strip())
    except sqlite3.OperationalError:
        pass
    try:
        for row in c.execute("SELECT DISTINCT proveedor FROM maestro_mps "
                             "WHERE proveedor IS NOT NULL AND proveedor != ''"):
            proveedores.add(row[0].strip())
    except sqlite3.OperationalError:
        pass
    # Sebastian 12-jun: incluir TAMBIEN el maestro de proveedores (tabla
    # proveedores) — antes solo salian los YA asignados a alguna MP/movimiento,
    # asi que para corregir/normalizar a un proveedor registrado pero aun no
    # usado (ej. Quincream es de GYM, no Agenquimicos) no aparecia en el
    # desplegable. Ahora salen TODOS los proveedores activos.
    try:
        for row in c.execute("SELECT DISTINCT nombre FROM proveedores "
                             "WHERE COALESCE(activo,1)=1 AND nombre IS NOT NULL AND nombre != ''"):
            proveedores.add(row[0].strip())
    except sqlite3.OperationalError:
        pass
    return jsonify({'proveedores': sorted(proveedores, key=lambda s: s.lower())})


def _normalizar_proveedor(nombre):
    """Normaliza nombre de proveedor para detección de duplicados.
    Sebastián 20-may-2026 · versión reforzada:
    - lowercase + strip + sin tildes (Ínchemical → inchemical)
    - quita sufijos jurídicos (SAS, LTDA, SA, SL, CIA, INC, CORP, LLC, BV, GMBH)
    - quita puntuación + caracteres especiales (. , ; : & - _ / \\)
    - colapsa espacios múltiples
    """
    import re, unicodedata
    n = (nombre or '').lower().strip()
    n = unicodedata.normalize('NFD', n)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    SUFIJOS = (
        r'\bs\.?a\.?s\.?\b', r'\bs\.?a\.?\b', r'\bltda\b', r'\bs\.?l\.?\b',
        r'\bcia\b', r'\binc\b', r'\bcorp(oration)?\b', r'\bllc\b',
        r'\b& cia\b', r'\bgmbh\b', r'\bbv\b', r'\bag\b', r'\bco\b',
        r'\bsrl\b', r'\bsac\b', r'\bspa\b',
    )
    for pat in SUFIJOS:
        n = re.sub(pat, '', n)
    n = re.sub(r'[.,;:&_/\\\-]+', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def _similitud_levenshtein(a, b):
    """Devuelve similitud 0-1 entre dos strings (1 = idénticos)."""
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    # Levenshtein iterativo
    if len(a) < len(b):
        a, b = b, a
    if len(b) == 0:
        return 0.0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        curr = [i + 1]
        for j, cb in enumerate(b):
            ins = prev[j + 1] + 1
            dele = curr[j] + 1
            subst = prev[j] + (0 if ca == cb else 1)
            curr.append(min(ins, dele, subst))
        prev = curr
    dist = prev[-1]
    max_len = max(len(a), len(b))
    return 1.0 - (dist / max_len)


# Tablas que tienen columna proveedor (o variante) · usadas por el
# unificador para reescribir referencias. Tupla: (tabla, columna).
# Sebastián 20-may-2026 · audit completo de schema.
_TABLAS_REF_PROVEEDOR = [
    ('movimientos', 'proveedor'),
    ('maestro_mps', 'proveedor'),
    ('maestro_mee', 'proveedor'),
    ('ordenes_compra', 'proveedor'),
    ('solicitudes_compra', 'proveedor'),
    ('solicitudes_compra_items', 'proveedor'),
    ('solicitudes_compra_items', 'proveedor_sugerido'),
    ('pagos_oc', 'proveedor'),
    ('mp_lead_time_config', 'proveedor_principal'),
    ('mee_lead_time_config', 'proveedor_default'),
    ('precios_mp_historico', 'proveedor'),
]


@bp.route('/api/proveedores-duplicados', methods=['GET'])
def proveedores_duplicados():
    """Agrupa proveedores que probablemente son el mismo con distinto
    formato. Sebastián 20-may-2026 · audit reforzado.

    Detección en 2 capas:
    1. Normalización exacta (lowercase, sin tildes, sin sufijos jurídicos,
       sin puntuación). Agrupa variantes que normalizan al mismo string.
    2. Similitud Levenshtein ≥ 0.85 entre las claves normalizadas ·
       captura typos (Inchemical vs Inquemical, BASF vs BAFS, etc.).

    Para cada grupo:
      - canonico_sugerido = variante con más caracteres
      - variantes ordenadas por uso (más usadas primero)
      - usos = dict con count de movimientos por variante
      - count_total_refs = suma de refs en TODAS las tablas

    Query params:
      ?similitud=0.85 (override threshold Levenshtein · 0=desactiva)
    """
    u, err, code = _require_session()
    if err:
        return err, code

    try:
        threshold = float(request.args.get('similitud', '0.85'))
        threshold = max(0.0, min(1.0, threshold))
    except (ValueError, TypeError):
        threshold = 0.85

    conn = get_db()
    c = conn.cursor()
    raw = set()
    # Cargar desde TODAS las tablas que tienen proveedor (no solo movs+maestro)
    for tabla, col in _TABLAS_REF_PROVEEDOR:
        try:
            for row in c.execute(
                f"SELECT DISTINCT {col} FROM {tabla} "
                f"WHERE {col} IS NOT NULL AND {col} != ''",
            ):
                v = (row[0] or '').strip()
                if v:
                    raw.add(v)
        except Exception:
            pass

    # Capa 1: agrupar por normalización exacta
    grupos_exacto = {}
    for nombre in raw:
        clave = _normalizar_proveedor(nombre)
        if not clave:
            continue
        grupos_exacto.setdefault(clave, []).append(nombre)

    # Capa 2: fusionar grupos con similitud ≥ threshold
    claves = list(grupos_exacto.keys())
    if threshold > 0:
        clusters = []
        usados = set()
        for i, k1 in enumerate(claves):
            if k1 in usados:
                continue
            cluster = {k1}
            usados.add(k1)
            for k2 in claves[i+1:]:
                if k2 in usados:
                    continue
                if _similitud_levenshtein(k1, k2) >= threshold:
                    cluster.add(k2)
                    usados.add(k2)
            clusters.append(cluster)
    else:
        clusters = [{k} for k in claves]

    # Construir grupos finales con stats
    duplicados = []
    for cluster_keys in clusters:
        variantes = []
        for k in cluster_keys:
            variantes.extend(grupos_exacto[k])
        if len(variantes) < 2:
            continue
        # Canonico sugerido: variante con más caracteres (más info)
        canonico = max(variantes, key=lambda s: (len(s), s))
        # Contar uso por variante en TODAS las tablas
        usos = {}
        refs_totales = {}
        for v in variantes:
            n_movs = 0
            n_refs_total = 0
            for tabla, col in _TABLAS_REF_PROVEEDOR:
                try:
                    n = c.execute(
                        f"SELECT COUNT(*) FROM {tabla} WHERE {col} = ?",
                        (v,)).fetchone()[0]
                    n_refs_total += int(n or 0)
                    if tabla == 'movimientos':
                        n_movs = int(n or 0)
                except Exception:
                    pass
            usos[v] = n_movs
            refs_totales[v] = n_refs_total
        duplicados.append({
            'clave_normalizada': ' | '.join(sorted(cluster_keys)),
            'canonico_sugerido': canonico,
            'variantes': sorted(variantes,
                                key=lambda s: -refs_totales.get(s, 0)),
            'usos': usos,
            'refs_totales': refs_totales,
            'count_variantes': len(variantes),
            'count_total_refs': sum(refs_totales.values()),
            'detectado_por': 'levenshtein' if len(cluster_keys) > 1 else 'normalizacion',
        })

    duplicados.sort(key=lambda g: (-g['count_total_refs'],
                                    -g['count_variantes']))
    return jsonify({
        'grupos': duplicados,
        'total_grupos': len(duplicados),
        'threshold_similitud': threshold,
        'tablas_consultadas': [t for t, _ in _TABLAS_REF_PROVEEDOR],
    })


@bp.route('/api/proveedores-unificar', methods=['POST'])
def proveedores_unificar():
    """Unifica varios alias de proveedor a un canonico.

    Body:
      canonico: str — nombre que queda como version oficial
      variantes: list[str] — todas las que se reemplazaran (incluye el
                              canonico tambien por idempotencia)

    Doble efecto (igual que editar proveedor de lote):
      1. UPDATE movimientos SET proveedor=canonico WHERE proveedor IN (variantes)
      2. UPDATE maestro_mps SET proveedor=canonico WHERE proveedor IN (variantes)

    Audit log con snapshot del cambio (variantes -> canonico, conteo).
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    canonico = (d.get('canonico') or '').strip()
    variantes = d.get('variantes') or []
    dry_run = bool(d.get('dry_run', False))
    if not canonico or len(canonico) < 2:
        return jsonify({'error': 'canonico requerido (>=2 chars)'}), 400
    if not isinstance(variantes, list) or len(variantes) < 1:
        return jsonify({'error': 'variantes requerido (lista no vacia)'}), 400
    # Limpiar variantes: dedup, strip, no vacias
    variantes = sorted({(v or '').strip() for v in variantes if (v or '').strip()})
    if not variantes:
        return jsonify({'error': 'variantes vacias tras limpieza'}), 400
    if len(variantes) > 50:
        return jsonify({'error': 'max 50 variantes por unificacion'}), 400

    conn = get_db()
    c = conn.cursor()
    placeholders = ','.join('?' * len(variantes))

    # Dry-run: contar filas que se actualizarían en cada tabla SIN tocar
    if dry_run:
        plan = {}
        total = 0
        for tabla, col in _TABLAS_REF_PROVEEDOR:
            try:
                n = c.execute(
                    f"SELECT COUNT(*) FROM {tabla} WHERE {col} IN ({placeholders})",
                    tuple(variantes),
                ).fetchone()[0]
                plan[f'{tabla}.{col}'] = int(n or 0)
                total += int(n or 0)
            except Exception as _e:
                plan[f'{tabla}.{col}'] = f'ERR: {str(_e)[:60]}'
        return jsonify({
            'dry_run': True,
            'canonico': canonico,
            'variantes': variantes,
            'plan_updates_por_tabla': plan,
            'total_filas_a_actualizar': total,
        })

    # APPLY · transaccional · reescribe en TODAS las tablas
    afectados = {}
    try:
        for tabla, col in _TABLAS_REF_PROVEEDOR:
            try:
                cur = c.execute(
                    f"UPDATE {tabla} SET {col} = ? "
                    f"WHERE {col} IN ({placeholders})",
                    (canonico, *variantes),
                )
                afectados[f'{tabla}.{col}'] = cur.rowcount or 0
            except Exception as _e:
                __import__('logging').getLogger('inventario').warning(
                    'unificar proveedor fallo tabla %s.%s: %s', tabla, col, _e)
                afectados[f'{tabla}.{col}'] = f'ERR: {str(_e)[:60]}'
        # audit_log master
        import json as _json
        c.execute(
            """INSERT INTO audit_log
                 (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
            (u, 'UNIFICAR_PROVEEDORES', '+'.join(t for t, _ in _TABLAS_REF_PROVEEDOR),
             canonico,
             _json.dumps({
                 'canonico': canonico,
                 'variantes_unificadas': variantes,
                 'filas_actualizadas_por_tabla': afectados,
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except Exception as _e:
        try: conn.rollback()
        except Exception: pass
        return jsonify({
            'error': f'transacción falló: {_e}',
            'afectados_parcial': afectados,
        }), 500

    total_actualizados = sum(
        v for v in afectados.values() if isinstance(v, int))
    return jsonify({
        'ok': True,
        'message': (f'Unificado a "{canonico}". '
                    f'{total_actualizados} filas actualizadas en '
                    f'{len(_TABLAS_REF_PROVEEDOR)} tablas.'),
        'canonico': canonico,
        'variantes_unificadas': variantes,
        'filas_actualizadas_por_tabla': afectados,
        'total_filas_actualizadas': total_actualizados,
        # Compat con UI vieja
        'movimientos_actualizados': afectados.get('movimientos.proveedor', 0)
            if isinstance(afectados.get('movimientos.proveedor'), int) else 0,
        'catalogo_actualizado': afectados.get('maestro_mps.proveedor', 0)
            if isinstance(afectados.get('maestro_mps.proveedor'), int) else 0,
    })


@bp.route('/api/lotes/<material_id>/<path:lote>/proveedor', methods=['PUT'])
def editar_proveedor_lote(material_id, lote):
    """Corrige el proveedor de un lote y del catálogo de la MP.

    Caso de uso (jefe de produccion): "este lote en realidad lo trajo
    Lyphar, no Inchemical — corrijamos para no repetir el bug en futuras
    recepciones".

    Doble efecto:
      1. UPDATE movimientos SET proveedor=? WHERE material_id=? AND lote=?
      2. UPDATE maestro_mps SET proveedor=? WHERE codigo_mp=?

    Esto evita que la siguiente recepción de la misma MP herede el
    proveedor erroneo desde el catalogo. El audit_log captura el cambio
    para trazabilidad (snapshot del valor anterior + nuevo).

    Body JSON:
      proveedor: str (obligatorio, 2..120 chars, no solo espacios)

    Soporta lote vacío con placeholder _SIN_LOTE_ (igual que DELETE).
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    nuevo_proveedor = (d.get('proveedor') or '').strip()
    if not nuevo_proveedor or len(nuevo_proveedor) < 2:
        return jsonify({
            'error': 'Proveedor invalido',
            'detail': 'Debe tener al menos 2 caracteres.'
        }), 400
    if len(nuevo_proveedor) > 120:
        return jsonify({
            'error': 'Proveedor demasiado largo',
            'detail': 'Maximo 120 caracteres.'
        }), 400

    sin_lote = (lote == '_SIN_LOTE_')

    conn = get_db()
    c = conn.cursor()

    # Snapshot valores actuales antes de cambiar
    if sin_lote:
        prov_anterior_row = c.execute(
            "SELECT MAX(proveedor) FROM movimientos "
            "WHERE material_id=? AND (lote IS NULL OR lote='')",
            (material_id,)
        ).fetchone()
    else:
        prov_anterior_row = c.execute(
            "SELECT MAX(proveedor) FROM movimientos "
            "WHERE material_id=? AND lote=?",
            (material_id, lote)
        ).fetchone()
    prov_anterior_lote = (prov_anterior_row[0] if prov_anterior_row else '') or ''

    cat_row = c.execute(
        "SELECT proveedor FROM maestro_mps WHERE codigo_mp=?",
        (material_id,)
    ).fetchone()
    if cat_row is None:
        return jsonify({
            'error': 'MP no encontrada',
            'detail': f'Codigo {material_id} no existe en maestro_mps'
        }), 404
    prov_anterior_cat = (cat_row[0] or '')

    # Update movimientos del lote
    if sin_lote:
        c.execute(
            "UPDATE movimientos SET proveedor=? "
            "WHERE material_id=? AND (lote IS NULL OR lote='')",
            (nuevo_proveedor, material_id)
        )
    else:
        c.execute(
            "UPDATE movimientos SET proveedor=? "
            "WHERE material_id=? AND lote=?",
            (nuevo_proveedor, material_id, lote)
        )
    movs_actualizados = c.rowcount

    # Update catalogo
    c.execute("UPDATE maestro_mps SET proveedor=? WHERE codigo_mp=?",
              (nuevo_proveedor, material_id))

    # Audit log
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (u, 'EDITAR_PROVEEDOR_LOTE', 'movimientos+maestro_mps',
                   f'{material_id}/{"" if sin_lote else lote}',
                   _json.dumps({
                       'material_id': material_id,
                       'lote': '' if sin_lote else lote,
                       'proveedor_anterior_lote': prov_anterior_lote,
                       'proveedor_anterior_catalogo': prov_anterior_cat,
                       'proveedor_nuevo': nuevo_proveedor,
                       'movimientos_actualizados': movs_actualizados,
                   }, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        __import__('logging').getLogger('inventario').warning(
            "audit_log no disponible — editar_proveedor por %s sobre %s/%s "
            "no quedo registrado.", u, material_id, lote,
        )

    conn.commit()

    return jsonify({
        'ok': True,
        'message': (f'Proveedor actualizado a "{nuevo_proveedor}" en '
                    f'{movs_actualizados} movimiento(s) del lote y en el '
                    f'catalogo de {material_id}.'),
        'movimientos_actualizados': movs_actualizados,
        'proveedor_anterior_lote': prov_anterior_lote,
        'proveedor_anterior_catalogo': prov_anterior_cat,
        'proveedor_nuevo': nuevo_proveedor,
    }), 200


@bp.route('/api/lotes/<material_id>/<path:lote>/ubicacion', methods=['PUT'])
def editar_ubicacion_lote(material_id, lote):
    """Corrige la ubicacion fisica (estanteria + posicion) de un lote.

    Sebastian 8-may-2026 (inventario REAL en vuelo): el modal de ajuste
    de Bodega MP no permitia cambiar posicion. Catalina/Sebastian
    encontraron discrepancias entre la posicion registrada y donde el
    lote esta fisicamente. Sin esto, el ajuste de cantidad funciona
    pero la pantalla sigue mostrando una posicion incorrecta.

    /api/lotes lee MAX(estanteria) y MAX(posicion) agrupado por
    (material_id, lote). UPDATE de TODOS los movimientos del lote
    deja el MAX en el valor nuevo y la UI lo refleja inmediatamente.

    Body JSON:
      estanteria: str (opcional, max 50 chars)
      posicion:   str (opcional, max 50 chars)
      motivo:     str (recomendado, queda en audit_log)

    Al menos uno de estanteria o posicion debe venir distinto del
    actual; si ambos coinciden, devuelve 200 sin tocar nada.

    Soporta lote vacio con placeholder _SIN_LOTE_ (igual que /proveedor).
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    nueva_estanteria = (d.get('estanteria') or '').strip()
    nueva_posicion = (d.get('posicion') or '').strip()
    motivo = (d.get('motivo') or '').strip()

    if len(nueva_estanteria) > 50:
        return jsonify({'error': 'Estanteria demasiado larga (max 50 chars)'}), 400
    if len(nueva_posicion) > 50:
        return jsonify({'error': 'Posicion demasiado larga (max 50 chars)'}), 400
    if not nueva_estanteria and not nueva_posicion:
        return jsonify({
            'error': 'Debe enviar al menos estanteria o posicion',
        }), 400

    sin_lote = (lote == '_SIN_LOTE_')

    conn = get_db()
    c = conn.cursor()

    # Snapshot valores actuales (MAX como en /api/lotes)
    if sin_lote:
        prev_row = c.execute(
            "SELECT MAX(estanteria), MAX(posicion), COUNT(*) "
            "FROM movimientos WHERE material_id=? AND (lote IS NULL OR lote='')",
            (material_id,)
        ).fetchone()
    else:
        prev_row = c.execute(
            "SELECT MAX(estanteria), MAX(posicion), COUNT(*) "
            "FROM movimientos WHERE material_id=? AND lote=?",
            (material_id, lote)
        ).fetchone()
    if not prev_row or (prev_row[2] or 0) == 0:
        return jsonify({
            'error': 'Lote no encontrado',
            'detail': f'No hay movimientos para {material_id}/{lote}',
        }), 404
    estanteria_anterior = (prev_row[0] or '')
    posicion_anterior = (prev_row[1] or '')

    # Determinar SET clause: solo updateamos los campos enviados
    sets = []
    params = []
    if nueva_estanteria:
        sets.append('estanteria=?')
        params.append(nueva_estanteria)
    if nueva_posicion:
        sets.append('posicion=?')
        params.append(nueva_posicion)
    if not sets:
        return jsonify({'ok': True, 'message': 'Nada que cambiar', 'movimientos_actualizados': 0}), 200

    where_sql = ("material_id=? AND (lote IS NULL OR lote='')"
                 if sin_lote else "material_id=? AND lote=?")
    where_params = [material_id] if sin_lote else [material_id, lote]

    sql = f"UPDATE movimientos SET {', '.join(sets)} WHERE {where_sql}"
    c.execute(sql, params + where_params)
    movs_actualizados = c.rowcount

    # Audit log con snapshot anterior+nuevo
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (u, 'EDITAR_UBICACION_LOTE', 'movimientos',
                   f'{material_id}/{"" if sin_lote else lote}',
                   _json.dumps({
                       'material_id': material_id,
                       'lote': '' if sin_lote else lote,
                       'estanteria_anterior': estanteria_anterior,
                       'posicion_anterior': posicion_anterior,
                       'estanteria_nueva': nueva_estanteria or estanteria_anterior,
                       'posicion_nueva': nueva_posicion or posicion_anterior,
                       'motivo': motivo,
                       'movimientos_actualizados': movs_actualizados,
                   }, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        __import__('logging').getLogger('inventario').warning(
            "audit_log no disponible — editar_ubicacion por %s sobre %s/%s "
            "no quedo registrado.", u, material_id, lote,
        )

    conn.commit()

    return jsonify({
        'ok': True,
        'message': (f'Ubicacion actualizada en {movs_actualizados} movimiento(s) '
                    f'del lote {lote if not sin_lote else "(sin lote)"} '
                    f'de {material_id}.'),
        'movimientos_actualizados': movs_actualizados,
        'estanteria_anterior': estanteria_anterior,
        'estanteria_nueva': nueva_estanteria or estanteria_anterior,
        'posicion_anterior': posicion_anterior,
        'posicion_nueva': nueva_posicion or posicion_anterior,
    }), 200


@bp.route('/api/lotes/<material_id>/<path:lote>/fecha-vencimiento', methods=['PUT'])
def editar_fecha_vencimiento_lote(material_id, lote):
    """Corrige la fecha de vencimiento de un lote.

    Sebastián 9-may-2026 (inventario REAL): "necesito poder modificar
    fecha de vencimiento si ves hay algunos que no tienen". Casos:
      - Lote ingresado sin fecha_vencimiento (Beauty Oil Copaiba en pantalla).
      - Fecha incorrecta por error de tipeo en recepción.
      - Lote con vencimiento extendido tras revisión QC.

    /api/lotes lee MAX(fecha_vencimiento) agrupado por (material_id, lote).
    UPDATE de TODOS los movimientos del lote deja MAX en valor nuevo y la
    UI lo refleja inmediato.

    Body JSON:
      fecha_vencimiento: str ISO YYYY-MM-DD (vacío = limpiar)
      motivo: str (recomendado, queda en audit_log)

    Validación: formato ISO. Si vacío, se permite (limpiar campo).
    Soporta lote vacío con placeholder _SIN_LOTE_.
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    nueva_fv = (d.get('fecha_vencimiento') or '').strip()
    motivo = (d.get('motivo') or '').strip()

    # Validar formato ISO si no vacío
    if nueva_fv:
        try:
            from datetime import date as _date
            _date.fromisoformat(nueva_fv[:10])
            nueva_fv = nueva_fv[:10]  # solo YYYY-MM-DD
        except (ValueError, TypeError):
            return jsonify({
                'error': 'fecha_vencimiento inválida',
                'detail': 'Formato esperado: YYYY-MM-DD',
            }), 400

    sin_lote = (lote == '_SIN_LOTE_')

    conn = get_db()
    c = conn.cursor()

    # Snapshot anterior
    if sin_lote:
        prev_row = c.execute(
            "SELECT MAX(fecha_vencimiento), COUNT(*) "
            "FROM movimientos WHERE material_id=? AND (lote IS NULL OR lote='')",
            (material_id,)
        ).fetchone()
    else:
        prev_row = c.execute(
            "SELECT MAX(fecha_vencimiento), COUNT(*) "
            "FROM movimientos WHERE material_id=? AND lote=?",
            (material_id, lote)
        ).fetchone()
    if not prev_row or (prev_row[1] or 0) == 0:
        return jsonify({
            'error': 'Lote no encontrado',
            'detail': f'No hay movimientos para {material_id}/{lote}',
        }), 404
    fv_anterior = (prev_row[0] or '')

    # Sin cambio
    if (fv_anterior or '')[:10] == (nueva_fv or '')[:10]:
        return jsonify({
            'ok': True,
            'message': 'Sin cambios respecto a la fecha actual',
            'movimientos_actualizados': 0,
            'fecha_anterior': fv_anterior,
            'fecha_nueva': nueva_fv,
        }), 200

    # UPDATE
    if sin_lote:
        c.execute(
            "UPDATE movimientos SET fecha_vencimiento=? "
            "WHERE material_id=? AND (lote IS NULL OR lote='')",
            (nueva_fv, material_id)
        )
    else:
        c.execute(
            "UPDATE movimientos SET fecha_vencimiento=? "
            "WHERE material_id=? AND lote=?",
            (nueva_fv, material_id, lote)
        )
    movs_actualizados = c.rowcount

    # Audit
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (u, 'EDITAR_FECHA_VENC_LOTE', 'movimientos',
                   f'{material_id}/{"" if sin_lote else lote}',
                   _json.dumps({
                       'material_id': material_id,
                       'lote': '' if sin_lote else lote,
                       'fecha_anterior': fv_anterior,
                       'fecha_nueva': nueva_fv,
                       'motivo': motivo,
                       'movimientos_actualizados': movs_actualizados,
                   }, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        __import__('logging').getLogger('inventario').warning(
            "audit_log no disponible — editar_fecha_venc por %s sobre %s/%s "
            "no quedo registrado.", u, material_id, lote,
        )

    conn.commit()

    return jsonify({
        'ok': True,
        'message': (f'Fecha de vencimiento actualizada en {movs_actualizados} '
                    f'movimiento(s) del lote {lote if not sin_lote else "(sin lote)"} '
                    f'de {material_id}.'),
        'movimientos_actualizados': movs_actualizados,
        'fecha_anterior': fv_anterior,
        'fecha_nueva': nueva_fv,
    }), 200


@bp.route('/api/lotes/<material_id>/<path:lote>/codigo-lote', methods=['PUT'])
def editar_codigo_lote(material_id, lote):
    """Renombra el número de lote de TODOS los movimientos de un lote.

    Sebastián 9-may-2026: "necesito que me deje cambiar lotes porque
    algunos estan mal". Caso típico: el lote se ingresó como '20250703'
    pero el formato correcto del proveedor es 'YT20250703'. Cambiar el
    código en todos los movimientos en una sola transacción.

    /api/lotes agrupa por (material_id, lote) → tras el rename, las
    filas del lote viejo desaparecen y aparece una nueva con el código
    correcto · trazabilidad histórica preservada en audit_log.

    Body JSON:
      lote_nuevo: str (1..120 chars, no espacios solo)
      motivo: str (recomendado, queda en audit_log)

    Validaciones:
      - lote_nuevo no vacío, distinto del actual.
      - lote_nuevo no debe colisionar: si ya existe un lote con ese
        código para el MISMO material_id, retorna 409 (mergearía
        dos lotes distintos · acción peligrosa que requiere flag
        explícito merge=true).
      - Si merge=true en body: permite la fusión (UPDATE igual aplica
        · stock se suma · audit registra la fusión).

    Soporta lote vacío con placeholder _SIN_LOTE_ (mismo patrón).
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    lote_nuevo = (d.get('lote_nuevo') or '').strip()
    motivo = (d.get('motivo') or '').strip()
    permitir_merge = bool(d.get('merge'))

    if not lote_nuevo:
        return jsonify({
            'error': 'lote_nuevo requerido',
            'detail': 'El nuevo número de lote no puede estar vacío.',
        }), 400
    if len(lote_nuevo) > 120:
        return jsonify({
            'error': 'lote_nuevo demasiado largo',
            'detail': 'Máximo 120 caracteres.',
        }), 400

    sin_lote = (lote == '_SIN_LOTE_')
    lote_actual = '' if sin_lote else lote

    # No-op si idéntico
    if lote_nuevo == lote_actual:
        return jsonify({
            'ok': True,
            'message': 'Sin cambios · el lote nuevo es igual al actual.',
            'movimientos_actualizados': 0,
        }), 200

    conn = get_db()
    c = conn.cursor()

    # Verificar que el lote actual existe
    if sin_lote:
        existe = c.execute(
            "SELECT COUNT(*) FROM movimientos "
            "WHERE material_id=? AND (lote IS NULL OR lote='')",
            (material_id,)
        ).fetchone()
    else:
        existe = c.execute(
            "SELECT COUNT(*) FROM movimientos WHERE material_id=? AND lote=?",
            (material_id, lote_actual)
        ).fetchone()
    n_actuales = (existe[0] if existe else 0) or 0
    if n_actuales == 0:
        return jsonify({
            'error': 'Lote no encontrado',
            'detail': f'No hay movimientos para {material_id}/{lote_actual or "(sin lote)"}',
        }), 404

    # Verificar colisión con lote_nuevo
    colision = c.execute(
        "SELECT COUNT(*) FROM movimientos WHERE material_id=? AND lote=?",
        (material_id, lote_nuevo)
    ).fetchone()
    n_colision = (colision[0] if colision else 0) or 0
    if n_colision > 0 and not permitir_merge:
        return jsonify({
            'error': 'Colisión de lote',
            'detail': (f'Ya existe un lote "{lote_nuevo}" para {material_id} '
                       f'con {n_colision} movimiento(s). Esto fusionaría dos '
                       f'lotes distintos en uno solo · acción peligrosa. '
                       f'Si querés fusionar deliberadamente, pasá merge=true '
                       f'en el body.'),
            'lote_existente_movs': n_colision,
            'lote_a_renombrar_movs': n_actuales,
        }), 409

    # Aplicar UPDATE
    if sin_lote:
        c.execute(
            "UPDATE movimientos SET lote=? "
            "WHERE material_id=? AND (lote IS NULL OR lote='')",
            (lote_nuevo, material_id)
        )
    else:
        c.execute(
            "UPDATE movimientos SET lote=? WHERE material_id=? AND lote=?",
            (lote_nuevo, material_id, lote_actual)
        )
    movs_actualizados = c.rowcount

    # Audit log con detalles del rename + fusión si aplica
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (u, 'EDITAR_CODIGO_LOTE', 'movimientos',
                   f'{material_id}/{lote_actual}',
                   _json.dumps({
                       'material_id': material_id,
                       'lote_anterior': lote_actual,
                       'lote_nuevo': lote_nuevo,
                       'motivo': motivo,
                       'movimientos_actualizados': movs_actualizados,
                       'fusion_realizada': permitir_merge and n_colision > 0,
                       'movs_lote_existente_pre_merge': n_colision,
                   }, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        __import__('logging').getLogger('inventario').warning(
            "audit_log no disponible — editar_codigo_lote por %s sobre "
            "%s/%s -> %s no quedó registrado.",
            u, material_id, lote_actual, lote_nuevo,
        )

    conn.commit()

    msg_extra = ''
    if permitir_merge and n_colision > 0:
        msg_extra = (f' · FUSIONADO con lote existente (sumó '
                     f'{n_colision} movimientos previos)')

    return jsonify({
        'ok': True,
        'message': (f'Lote renombrado de "{lote_actual or "(sin lote)"}" a '
                    f'"{lote_nuevo}" en {movs_actualizados} movimiento(s) '
                    f'de {material_id}.{msg_extra}'),
        'movimientos_actualizados': movs_actualizados,
        'lote_anterior': lote_actual,
        'lote_nuevo': lote_nuevo,
        'fusion_realizada': permitir_merge and n_colision > 0,
    }), 200


@bp.route('/api/lotes/<material_id>/<path:lote>', methods=['DELETE'])
def eliminar_lote(material_id, lote):
    """Elimina un lote completo por incoherencia (jefe de produccion).

    Hard-delete de todos los movimientos que comparten (material_id, lote).
    El bloque ENTRADA original + sus salidas asociadas se borran. Antes de
    borrar, se hace snapshot al audit_log con el estado completo del lote
    (cantidad neta, fecha venc, proveedor, # movs) + el motivo del usuario,
    para mantener trazabilidad de lo que se borro y por que.

    Body JSON:
      motivo: str (obligatorio, min 10 chars) - razon documentada

    Caso de uso: recepcion duplicada, codigo MP equivocado, lote registrado
    contra el material erroneo, etc. Para correcciones de cantidad usar
    'Ajustar' en su lugar (genera contra-movimiento, preserva historia).
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    motivo = (d.get('motivo') or '').strip()
    if len(motivo) < 10:
        return jsonify({
            'error': 'Motivo obligatorio',
            'detail': 'Explica por que eliminas este lote (min 10 caracteres). '
                      'Esto queda en audit_log para trazabilidad.'
        }), 400

    # _SIN_LOTE_ es placeholder del frontend para movimientos sin lote
    # (lote NULL o ''). Hacemos match contra ambos.
    sin_lote = (lote == '_SIN_LOTE_')

    conn = get_db()
    c = conn.cursor()

    # Snapshot del lote antes de borrar — solo columnas garantizadas en
    # el schema actual de movimientos (ver database.py CREATE TABLE).
    cols_select = ("m.id, m.tipo, m.cantidad, m.fecha, m.proveedor, "
                   "m.fecha_vencimiento, m.operador, m.observaciones")
    if sin_lote:
        c.execute(f"""SELECT {cols_select}
                     FROM movimientos m
                     WHERE m.material_id=? AND (m.lote IS NULL OR m.lote='')
                     ORDER BY m.fecha ASC""", (material_id,))
    else:
        c.execute(f"""SELECT {cols_select}
                     FROM movimientos m
                     WHERE m.material_id=? AND m.lote=?
                     ORDER BY m.fecha ASC""", (material_id, lote))
    movs = c.fetchall()
    if not movs:
        return jsonify({
            'error': 'Lote no encontrado',
            'detail': f'No existen movimientos para {material_id} / {lote}'
        }), 404

    # Calcular saldo neto + nombre comercial para el log.
    # Ajuste=resta fix 28-may · antes todo lo no-'Entrada' se restaba, así un
    # 'Ajuste'/'Ajuste +' positivo se contaba negativo y el contra-movimiento
    # quedaba mal (dejaba stock residual/duplicado). Patrón canónico de signo.
    def _signo_mov(tipo):
        t = tipo or ''
        if t in ('Entrada', 'entrada', 'ENTRADA', 'Ajuste +', 'Ajuste'):
            return 1
        if t in ('Salida', 'salida', 'SALIDA', 'Ajuste -'):
            return -1
        return 0
    saldo_neto = sum(_signo_mov(mv[1]) * (mv[2] or 0) for mv in movs)
    nombre_row = c.execute(
        "SELECT nombre_comercial FROM maestro_mps WHERE codigo_mp=?",
        (material_id,)
    ).fetchone()
    nombre_comercial = (nombre_row[0] if nombre_row else '') or material_id

    snapshot = {
        'material_id': material_id,
        'nombre_comercial': nombre_comercial,
        'lote': '' if sin_lote else lote,
        'saldo_neto_g_al_eliminar': round(saldo_neto, 2),
        'num_movimientos': len(movs),
        'fechas': [str(mv[3])[:10] for mv in movs if mv[3]],
        'proveedores': sorted({(mv[4] or '') for mv in movs if mv[4]}),
        'operadores':  sorted({(mv[6] or '') for mv in movs if mv[6]}),
        'motivo':      motivo,
    }

    # Audit log antes de borrar (sobrevive aunque algo falle abajo)
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (u, 'ELIMINAR_LOTE', 'movimientos',
                   f'{material_id}/{lote}',
                   _json.dumps(snapshot, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        # audit_log no existe en deploy super-viejo — log en Sentry
        # via excepcion explicita y seguir.
        __import__('logging').getLogger('inventario').warning(
            "audit_log no disponible — eliminar_lote por %s sobre %s/%s "
            "no quedo registrado en BD. motivo=%s",
            u, material_id, lote, motivo,
        )

    # FIX-B5 13-may-2026: ANTES hacíamos hard DELETE de TODOS los movs del
    # lote (Entrada + Salidas FEFO consumidas en producciones reales) ·
    # rompe trazabilidad INVIMA Resolución 2214/2021 art. 10 (qué producción
    # consumió este lote desaparecía del kardex). El agregado por material_id
    # quedaba consistente pero el detalle por lote se borraba.
    # Ahora: contra-movimiento del saldo neto (Salida si saldo>0 · Entrada si
    # saldo<0). Las filas históricas de Entrada y Salidas FEFO se preservan.
    if abs(saldo_neto) > 0.01:
        tipo_contra = 'Salida' if saldo_neto > 0 else 'Entrada'
        obs_contra = (f'[ELIMINAR_LOTE] Saldo neto {saldo_neto:.2f}g cancelado '
                      f'por eliminación de lote {lote}. Motivo: {motivo[:200]}')
        c.execute("""INSERT INTO movimientos
                       (material_id, material_nombre, cantidad, tipo, fecha,
                        observaciones, operador, lote, estado_lote)
                     VALUES (?, ?, ?, ?, datetime('now', '-5 hours', 'utc'), ?, ?, ?, 'ELIMINADO')""",
                  (material_id, nombre_comercial, abs(saldo_neto), tipo_contra,
                   obs_contra, u, lote if not sin_lote else ''))
    deleted = 0  # ya no hay DELETE · histórico preservado
    conn.commit()

    return jsonify({
        'ok': True,
        'message': (f'Lote {lote} de {nombre_comercial} eliminado (saldo neto '
                    f'{saldo_neto:.2f}g cancelado con contra-movimiento). '
                    f'Histórico preservado para trazabilidad INVIMA.'),
        'saldo_cancelado_g': saldo_neto,
        'deleted_count': deleted,
        'snapshot': snapshot,
    }), 200


@bp.route('/api/admin/auditar-formula-mp-match', methods=['GET'])
def auditar_formula_mp_match():
    """Sebastián 22-may-2026: auditar TODAS las fórmulas vs maestro MPs.

    Detecta materiales en formula_items que NO matchean con ningún MP en
    maestro_mps (huérfanos). Para cada huérfano:
      - Intenta expansión vía mp_aliases (SAP→Sodium Ascorbyl Phosphate)
      - Busca MP con INCI canonical · si existe sugiere fix
      - Si NO existe MP · sugiere crearlo

    Returns: {
      huerfanos: [{material_id, material_nombre, productos_usando[], sugerencia}],
      conteo: {total_formulas, huerfanos_count, auto_reparables, requiere_crear_mp},
      aliases_usados: int,
    }
    """
    u, err, code = _require_admin()
    if err: return err, code
    conn = get_db(); c = conn.cursor()

    # 1. Set de TODOS los códigos + nombres en maestro_mps
    mp_codigos = set()
    mp_por_nombre = {}  # LOWER(nombre) → codigo_mp
    mp_por_inci = {}    # LOWER(inci) → codigo_mp
    try:
        rows_mp = c.execute(
            """SELECT codigo_mp, COALESCE(nombre_comercial,''),
                      COALESCE(nombre_inci,''), COALESCE(activo,1)
               FROM maestro_mps"""
        ).fetchall()
    except Exception:
        rows_mp = []
    for cod, nom_c, nom_i, activo in rows_mp:
        mp_codigos.add(cod)
        if nom_c:
            mp_por_nombre[nom_c.lower().strip()] = (cod, activo)
        if nom_i:
            mp_por_inci[nom_i.lower().strip()] = (cod, activo)

    # 2. Aliases conocidos
    aliases_dict = {}  # LOWER(alias) → INCI canonical
    try:
        rows_al = c.execute(
            """SELECT alias, nombre_inci_canonical
               FROM mp_aliases WHERE COALESCE(activo,1)=1"""
        ).fetchall()
        for al, inci in rows_al:
            aliases_dict[al.lower().strip()] = inci
    except Exception:
        pass

    # 3. Materiales únicos en formula_items
    huerfanos = {}
    auto_reparables = 0
    requiere_crear = 0
    try:
        rows_form = c.execute(
            """SELECT material_id, material_nombre, producto_nombre, COUNT(*) as usos
               FROM formula_items
               WHERE COALESCE(material_nombre,'')!='' OR COALESCE(material_id,'')!=''
               GROUP BY material_id, material_nombre, producto_nombre"""
        ).fetchall()
    except Exception:
        rows_form = []

    materiales_unicos = {}  # (mat_id, mat_nom) → {productos: [], usos: int}
    for mat_id, mat_nom, producto, usos in rows_form:
        key = ((mat_id or '').strip(), (mat_nom or '').strip())
        if key not in materiales_unicos:
            materiales_unicos[key] = {'productos': [], 'usos': 0}
        materiales_unicos[key]['productos'].append(producto)
        materiales_unicos[key]['usos'] += int(usos or 0)

    total_formulas = len(materiales_unicos)

    for (mat_id, mat_nom), data in materiales_unicos.items():
        # Match por codigo
        if mat_id and mat_id in mp_codigos:
            continue
        # Match por nombre comercial (exact lowercase)
        nom_lower = mat_nom.lower().strip()
        match_por_nombre = mp_por_nombre.get(nom_lower) or mp_por_inci.get(nom_lower)
        if match_por_nombre:
            cod_real, activo_real = match_por_nombre
            if activo_real:
                continue
            # Existe pero inactivo
            huerfanos[(mat_id, mat_nom)] = {
                'material_id': mat_id,
                'material_nombre': mat_nom,
                'productos': data['productos'][:5],
                'usos_total': data['usos'],
                'estado': 'MP_INACTIVO',
                'sugerencia': f'MP existe ({cod_real}) pero activo=0 · reactivar',
                'accion': 'reactivar',
                'codigo_sugerido': cod_real,
            }
            continue
        # Match parcial INCI o comercial (contains)
        sugerencia_match = None
        for inci_lower, (cod_i, act_i) in mp_por_inci.items():
            if nom_lower in inci_lower or inci_lower in nom_lower:
                sugerencia_match = (cod_i, inci_lower, 'inci_parcial')
                break
        if not sugerencia_match:
            for nom2_lower, (cod_n, act_n) in mp_por_nombre.items():
                if nom_lower in nom2_lower or nom2_lower in nom_lower:
                    sugerencia_match = (cod_n, nom2_lower, 'nombre_parcial')
                    break

        # Match por alias
        alias_inci = aliases_dict.get(nom_lower)
        if alias_inci:
            # Buscar MP con ese INCI
            alias_inci_lower = alias_inci.lower().strip()
            mp_via_alias = mp_por_inci.get(alias_inci_lower)
            if not mp_via_alias:
                # parcial
                for inci_lower, (cod_i, act_i) in mp_por_inci.items():
                    if alias_inci_lower in inci_lower:
                        mp_via_alias = (cod_i, act_i)
                        break
            if mp_via_alias:
                cod_via_alias, act_via = mp_via_alias
                huerfanos[(mat_id, mat_nom)] = {
                    'material_id': mat_id,
                    'material_nombre': mat_nom,
                    'productos': data['productos'][:5],
                    'usos_total': data['usos'],
                    'estado': 'ALIAS_DETECTADO',
                    'alias_expandido': alias_inci,
                    'sugerencia': f'Abreviatura {mat_nom} = {alias_inci} → MP {cod_via_alias}',
                    'accion': 'renombrar_y_vincular',
                    'codigo_sugerido': cod_via_alias,
                    'nombre_sugerido': alias_inci,
                }
                auto_reparables += 1
                continue
            else:
                # Alias detectado pero MP no existe
                huerfanos[(mat_id, mat_nom)] = {
                    'material_id': mat_id,
                    'material_nombre': mat_nom,
                    'productos': data['productos'][:5],
                    'usos_total': data['usos'],
                    'estado': 'ALIAS_SIN_MP',
                    'alias_expandido': alias_inci,
                    'sugerencia': f'{mat_nom} = {alias_inci} pero NO existe MP con ese INCI · CREAR',
                    'accion': 'crear_mp',
                    'nombre_sugerido': alias_inci,
                }
                requiere_crear += 1
                continue
        if sugerencia_match:
            cod_s, nom_s, tipo_s = sugerencia_match
            huerfanos[(mat_id, mat_nom)] = {
                'material_id': mat_id,
                'material_nombre': mat_nom,
                'productos': data['productos'][:5],
                'usos_total': data['usos'],
                'estado': 'MATCH_PARCIAL',
                'sugerencia': f'Match parcial ({tipo_s}) con {cod_s} ({nom_s})',
                'accion': 'verificar',
                'codigo_sugerido': cod_s,
            }
            continue
        # No match en nada
        huerfanos[(mat_id, mat_nom)] = {
            'material_id': mat_id,
            'material_nombre': mat_nom,
            'productos': data['productos'][:5],
            'usos_total': data['usos'],
            'estado': 'SIN_MATCH',
            'sugerencia': f'NO existe en maestro_mps · ¿crear o agregar alias?',
            'accion': 'crear_mp_o_alias',
        }
        requiere_crear += 1

    huerfanos_list = sorted(huerfanos.values(),
                            key=lambda x: (-x['usos_total'], x['material_nombre']))

    return jsonify({
        'huerfanos': huerfanos_list,
        'conteo': {
            'total_materiales_formula': total_formulas,
            'huerfanos': len(huerfanos),
            'porcentaje_huerfanos': round(len(huerfanos) / max(total_formulas, 1) * 100, 1),
            'auto_reparables_alias': auto_reparables,
            'requiere_crear_mp': requiere_crear,
        },
        'aliases_cargados': len(aliases_dict),
    })


@bp.route('/api/admin/normalizar-formulas-mp', methods=['POST'])
def normalizar_formulas_mp():
    """Sebastián 22-may-2026 · aplica fix automático a fórmulas con abbreviaturas.

    Body: {
      dry_run: bool (default True) · si False realmente actualiza
      solo_alias: bool (default True) · solo normaliza casos con alias detectado
    }

    Para cada material en formula_items con alias detectado:
      - UPDATE material_nombre = alias.nombre_inci_canonical
      - UPDATE material_id = código del MP que matchea ese INCI
      - audit_log NORMALIZAR_FORMULA_MP por cada fila
    """
    u, err, code = _require_admin()
    if err: return err, code
    d = request.json or {}
    dry_run = d.get('dry_run', True)
    solo_alias = d.get('solo_alias', True)

    conn = get_db(); c = conn.cursor()

    # Cargar aliases + MPs (mismo prep que auditoría)
    aliases_dict = {}
    try:
        for al, inci in c.execute(
            "SELECT alias, nombre_inci_canonical FROM mp_aliases WHERE COALESCE(activo,1)=1"
        ).fetchall():
            aliases_dict[al.lower().strip()] = inci
    except Exception:
        pass
    mp_por_inci = {}
    try:
        for cod, inci, activo in c.execute(
            "SELECT codigo_mp, COALESCE(nombre_inci,''), COALESCE(activo,1) FROM maestro_mps"
        ).fetchall():
            if inci and activo:
                mp_por_inci[inci.lower().strip()] = (cod, inci)
    except Exception:
        pass

    # Lista materiales únicos
    rows_form = c.execute(
        """SELECT DISTINCT material_id, material_nombre FROM formula_items
           WHERE COALESCE(material_nombre,'')!=''"""
    ).fetchall()

    cambios = []
    for mat_id, mat_nom in rows_form:
        nom_lower = (mat_nom or '').lower().strip()
        # Solo procesar si es abreviatura conocida
        alias_inci = aliases_dict.get(nom_lower)
        if not alias_inci:
            continue
        alias_lower = alias_inci.lower().strip()
        mp_match = mp_por_inci.get(alias_lower)
        if not mp_match:
            # parcial
            for k, v in mp_por_inci.items():
                if alias_lower in k:
                    mp_match = v
                    break
        if not mp_match:
            continue
        cod_new, nom_new = mp_match
        cambios.append({
            'mat_id_antes': mat_id,
            'mat_nom_antes': mat_nom,
            'mat_id_despues': cod_new,
            'mat_nom_despues': nom_new,
        })
        if not dry_run:
            try:
                c.execute(
                    """UPDATE formula_items
                       SET material_id=?, material_nombre=?
                       WHERE COALESCE(material_id,'')=? AND material_nombre=?""",
                    (cod_new, nom_new, mat_id or '', mat_nom),
                )
                try:
                    audit_log(c, usuario=u, accion='NORMALIZAR_FORMULA_MP',
                              tabla='formula_items', registro_id=f'{mat_id}/{mat_nom}',
                              antes={'material_id': mat_id, 'material_nombre': mat_nom},
                              despues={'material_id': cod_new, 'material_nombre': nom_new})
                except Exception:
                    pass
            except Exception as e:
                log.warning('normalizar fila fallo %s/%s: %s', mat_id, mat_nom, e)
    if not dry_run:
        conn.commit()

    return jsonify({
        'ok': True,
        'dry_run': dry_run,
        'cambios_aplicados' if not dry_run else 'cambios_propuestos': len(cambios),
        'detalle': cambios[:50],
        'mensaje': (
            f'{len(cambios)} fórmulas {"normalizadas" if not dry_run else "se normalizarían"} · '
            f'pasá dry_run=false para aplicar' if dry_run else ''
        ),
    })


@bp.route('/api/maestro-mps/export-lista-simple', methods=['GET'])
def maestro_mps_export_lista_simple():
    """Exporta lista simple de materias primas · Sebastián 19-may-2026.

    Alejandro pidió "una lista de Excel solo con las materias primas, sin
    precio, sin proveedor, sin lo que tenemos, solo la lista".

    Formato (querystring `?fmt=`):
      - `xlsx` (default) · archivo Excel nativo · header en negrita,
        columnas auto-ajustadas, filtro automático. Lo más cómodo para
        Alejandro (Excel en español rompe CSVs con coma).
      - `csv` · CSV con `;` (Excel-ES) + BOM UTF-8. Fallback liviano.

    Solo MPs activas (activo=1), ordenadas por nombre.
    """
    from flask import make_response
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    fmt = (request.args.get('fmt') or 'xlsx').lower().strip()

    conn = get_db()
    c = conn.cursor()
    rows = c.execute(
        """SELECT COALESCE(codigo_mp,'') AS codigo,
                  COALESCE(nombre_comercial,'') AS nom_com,
                  COALESCE(nombre_inci,'') AS nom_inci,
                  COALESCE(tipo_material, tipo, 'MP') AS tipo
           FROM maestro_mps
           WHERE COALESCE(activo, 1) = 1
           ORDER BY COALESCE(nombre_comercial, nombre_inci, codigo_mp) ASC""",
    ).fetchall()
    from datetime import datetime as _dt
    fecha_str = _dt.now().strftime('%Y-%m-%d')

    HEADERS = ['Codigo', 'Nombre Comercial', 'Nombre INCI', 'Tipo']

    if fmt == 'csv':
        # CSV con `;` (Excel-ES) + BOM UTF-8
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        buf.write('﻿')  # BOM UTF-8
        writer = _csv.writer(buf, delimiter=';', quoting=_csv.QUOTE_MINIMAL)
        writer.writerow(HEADERS)
        for r in rows:
            writer.writerow([r[0], r[1], r[2], r[3]])
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = (
            f'attachment; filename="materias-primas-{fecha_str}.csv"')
        resp.headers['X-Content-Type-Options'] = 'nosniff'
        return resp

    # XLSX nativo · default
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = 'Materias Primas'
    # Header con formato
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0F766E', end_color='0F766E',
                              fill_type='solid')
    center = Alignment(horizontal='left', vertical='center')
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    # Filas
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, val in enumerate(r, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # Auto-ajustar ancho columnas (heurística por largo máximo)
    for col_idx, _h in enumerate(HEADERS, start=1):
        letra = get_column_letter(col_idx)
        max_len = len(HEADERS[col_idx - 1])
        for r_idx in range(2, len(rows) + 2):
            v = ws.cell(row=r_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letra].width = min(max_len + 2, 60)
    # Congelar header + filtro auto
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = (
        f'attachment; filename="materias-primas-{fecha_str}.xlsx"')
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    return resp


@bp.route('/api/maestro-mps/next-codigo', methods=['GET'])
def maestro_mps_next_codigo():
    """Devuelve el SIGUIENTE código MP disponible.

    Sebastián 8-may-2026: el form "Crear Nueva MP" pedía al usuario
    escribir manualmente el código (MP00350) sin saber cuál era el
    último. Resultado: códigos repetidos o saltos al azar.

    Lógica:
      1. Busca todos los códigos que matchean el patrón MP\\d+
      2. Encuentra el número máximo
      3. Devuelve max+1 con el mismo padding (5 dígitos = MP00350)

    Si no existe ninguna MP en el catálogo, arranca en MP00001.

    Returns:
      { siguiente: 'MP00350', ultimo: 'MP00349', total: 248 }

    Querystring:
      ?prefix=MP  · permite cambiar prefijo (default 'MP')
      ?width=5    · padding del número (default 5)
    """
    u, err, code = _require_session()
    if err:
        return err, code

    prefix = (request.args.get('prefix') or 'MP').strip().upper()
    try:
        width = max(3, min(int(request.args.get('width', 5)), 10))
    except (ValueError, TypeError):
        width = 5

    conn = get_db(); c = conn.cursor()
    # Pattern SQL: MP seguido de solo dígitos
    pattern = f'{prefix}%'
    rows = c.execute(
        "SELECT codigo_mp FROM maestro_mps WHERE codigo_mp LIKE ?", (pattern,)
    ).fetchall()

    import re as _re
    regex = _re.compile(f'^{_re.escape(prefix)}(\\d+)$')
    max_n = 0
    ultimo = None
    total_validos = 0
    for (cod,) in rows:
        if not cod:
            continue
        m = regex.match(cod.strip().upper())
        if not m:
            continue
        try:
            n = int(m.group(1))
            total_validos += 1
            if n > max_n:
                max_n = n
                ultimo = cod
        except (ValueError, TypeError):
            continue

    siguiente_n = max_n + 1
    siguiente = f'{prefix}{siguiente_n:0{width}d}'

    return jsonify({
        'siguiente': siguiente,
        'ultimo': ultimo,
        'siguiente_n': siguiente_n,
        'total_con_prefix': total_validos,
        'total_en_catalogo': len(rows),
        'prefix': prefix,
        'width': width,
    })


@bp.route('/api/maestro-mps/buscar-inteligente', methods=['GET'])
def maestro_mps_buscar_inteligente():
    """SHOPIFY-FIX · 22-may-2026 · búsqueda inteligente por abreviatura INCI.

    Sebastián 22-may-2026: 'fórmula MAXLASH dice SAP es Sodium Ascorbyl Phosphate
    · cuando busco SAP en inventario no me sale'.

    Estrategia:
      1. Query exacta en codigo_mp, nombre_comercial, nombre_inci (LIKE)
      2. Si la query matchea un alias (mp_aliases.alias case-insensitive) ·
         expandir a INCI canonical + buscar por ese INCI
      3. Normalización: sin acentos, sin guiones, lowercase
      4. Devolver ranked por relevancia (exact > startswith > contains)

    Query params:
      q: string mínimo 2 chars
      limit: max 20 (default 12)
      include_inactive: '1' incluye activo=0

    Returns: {mps: [{codigo_mp, nombre_inci, nombre_comercial, ...,
                     match_via: 'codigo'|'comercial'|'inci'|'alias'|'norm',
                     match_score: 0-100}], aliases_aplicados: [...]}
    """
    u, err, code = _require_session()
    if err: return err, code
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'mps': [], 'mensaje': 'Query mínimo 2 chars'})
    try:
        limit = max(1, min(int(request.args.get('limit', 12)), 50))
    except (TypeError, ValueError):
        limit = 12
    include_inactive = request.args.get('include_inactive') == '1'

    import unicodedata as _ud, re as _re
    def _norm(s):
        if not s:
            return ''
        s2 = _ud.normalize('NFD', str(s)).encode('ascii', 'ignore').decode('ascii')
        return _re.sub(r'[\-_\.\s]+', ' ', s2.lower()).strip()

    q_norm = _norm(q)
    conn = get_db(); c = conn.cursor()

    # 1. Aliases · expandir abreviatura a INCI canonical
    aliases_aplicados = []
    inci_expansion = []
    try:
        alias_rows = c.execute(
            """SELECT alias, codigo_mp, nombre_inci_canonical, tipo
               FROM mp_aliases
               WHERE LOWER(alias)=? AND COALESCE(activo,1)=1""",
            (q.lower(),),
        ).fetchall()
        for ar in alias_rows:
            aliases_aplicados.append({
                'alias': ar[0], 'inci_canonical': ar[2], 'tipo': ar[3],
            })
            if ar[2]:
                inci_expansion.append(ar[2])
    except Exception:
        pass

    # 2. Búsqueda principal
    where_activo = '' if include_inactive else 'AND activo=1 '
    # Q literal
    rows = c.execute(
        f"""SELECT codigo_mp, COALESCE(nombre_inci,''), COALESCE(nombre_comercial,''),
                   COALESCE(tipo,''), COALESCE(proveedor,''),
                   COALESCE(stock_minimo,0), COALESCE(precio_referencia,0),
                   COALESCE(tipo_material,'MP'), activo
            FROM maestro_mps
            WHERE 1=1 {where_activo}
              AND (LOWER(codigo_mp) LIKE ?
                   OR LOWER(COALESCE(nombre_comercial,'')) LIKE ?
                   OR LOWER(COALESCE(nombre_inci,'')) LIKE ?)
            LIMIT 200""",
        (f'%{q.lower()}%', f'%{q.lower()}%', f'%{q.lower()}%'),
    ).fetchall()

    matches = {}
    for r in rows:
        cod = r[0]
        score = 0
        via = 'norm'
        if q.lower() in cod.lower():
            score = 100; via = 'codigo'
        elif q.lower() in (r[2] or '').lower():
            score = 80; via = 'comercial'
        elif q.lower() in (r[1] or '').lower():
            score = 70; via = 'inci'
        matches[cod] = (score, via, r)

    # 3. Expansión vía alias (búsqueda extra)
    for inci_can in inci_expansion:
        rows_alias = c.execute(
            f"""SELECT codigo_mp, COALESCE(nombre_inci,''), COALESCE(nombre_comercial,''),
                       COALESCE(tipo,''), COALESCE(proveedor,''),
                       COALESCE(stock_minimo,0), COALESCE(precio_referencia,0),
                       COALESCE(tipo_material,'MP'), activo
                FROM maestro_mps
                WHERE 1=1 {where_activo}
                  AND (LOWER(COALESCE(nombre_inci,'')) LIKE ?
                       OR LOWER(COALESCE(nombre_comercial,'')) LIKE ?)
                LIMIT 50""",
            (f'%{inci_can.lower()}%', f'%{inci_can.lower()}%'),
        ).fetchall()
        for r in rows_alias:
            cod = r[0]
            if cod not in matches:
                matches[cod] = (90, 'alias', r)

    # 4. Búsqueda por normalización (sin acentos, sin guiones)
    if len(matches) < limit:
        try:
            # Lista todos los MPs y filtro en Python (no hay func normalize en SQLite)
            all_rows = c.execute(
                f"""SELECT codigo_mp, COALESCE(nombre_inci,''), COALESCE(nombre_comercial,''),
                           COALESCE(tipo,''), COALESCE(proveedor,''),
                           COALESCE(stock_minimo,0), COALESCE(precio_referencia,0),
                           COALESCE(tipo_material,'MP'), activo
                    FROM maestro_mps
                    WHERE 1=1 {where_activo}
                    LIMIT 5000""",
            ).fetchall()
            for r in all_rows:
                if r[0] in matches:
                    continue
                if (q_norm in _norm(r[2]) or q_norm in _norm(r[1])):
                    matches[r[0]] = (50, 'norm', r)
                    if len(matches) >= limit + 5:
                        break
        except Exception:
            pass

    # Ranking + format
    sorted_matches = sorted(matches.values(), key=lambda x: -x[0])[:limit]
    out = []
    for score, via, r in sorted_matches:
        out.append({
            'codigo_mp': r[0],
            'nombre_inci': r[1],
            'nombre_comercial': r[2],
            'tipo': r[3],
            'proveedor': r[4],
            'stock_minimo': r[5],
            'precio_referencia': r[6],
            'tipo_material': r[7],
            'activo': bool(r[8]),
            'match_via': via,
            'match_score': score,
        })
    return jsonify({
        'mps': out,
        'total': len(out),
        'query': q,
        'aliases_aplicados': aliases_aplicados,
    })


@bp.route('/api/maestro-mps/alias', methods=['GET', 'POST', 'DELETE'])
def maestro_mps_alias():
    """Sebastián 22-may-2026 · CRUD aliases de MP · agregar abreviaturas
    cuando aparecen casos nuevos · auto-aprendizaje.

    GET ?q=SAP → busca aliases por substring
    POST {alias, codigo_mp?, nombre_inci_canonical, tipo} → INSERT
    DELETE ?id=N → remove
    """
    u, err, code = _require_session()
    if err: return err, code
    conn = get_db(); c = conn.cursor()
    if request.method == 'GET':
        q = (request.args.get('q') or '').strip()
        sql = "SELECT id, alias, codigo_mp, nombre_inci_canonical, tipo, fuente, creado_por, creado_en, activo FROM mp_aliases WHERE COALESCE(activo,1)=1"
        params = []
        if q:
            sql += " AND (LOWER(alias) LIKE ? OR LOWER(nombre_inci_canonical) LIKE ?)"
            params.extend([f'%{q.lower()}%', f'%{q.lower()}%'])
        sql += " ORDER BY alias LIMIT 100"
        rows = c.execute(sql, params).fetchall()
        return jsonify({'aliases': [
            {'id': r[0], 'alias': r[1], 'codigo_mp': r[2],
             'nombre_inci_canonical': r[3], 'tipo': r[4], 'fuente': r[5],
             'creado_por': r[6], 'creado_en': r[7], 'activo': bool(r[8])}
            for r in rows
        ]})
    if request.method == 'POST':
        d = request.json or {}
        alias = (d.get('alias') or '').strip()
        inci = (d.get('nombre_inci_canonical') or '').strip()
        if not alias or not inci:
            return jsonify({'error': 'alias y nombre_inci_canonical requeridos'}), 400
        tipo = d.get('tipo', 'abreviatura')
        if tipo not in ('abreviatura', 'sinonimo', 'typo_comun', 'translation'):
            tipo = 'abreviatura'
        codigo_mp = d.get('codigo_mp') or None
        try:
            c.execute(
                """INSERT INTO mp_aliases (alias, codigo_mp, nombre_inci_canonical,
                                            tipo, fuente, creado_por)
                   VALUES (?, ?, ?, ?, 'manual', ?)""",
                (alias, codigo_mp, inci, tipo, u),
            )
            conn.commit()
            return jsonify({'ok': True, 'id': c.lastrowid})
        except Exception as e:
            return jsonify({'error': str(e)}), 400
    if request.method == 'DELETE':
        try:
            aid = int(request.args.get('id', 0))
        except (TypeError, ValueError):
            return jsonify({'error': 'id inválido'}), 400
        if not aid:
            return jsonify({'error': 'id requerido'}), 400
        c.execute("UPDATE mp_aliases SET activo=0 WHERE id=?", (aid,))
        conn.commit()
        return jsonify({'ok': True})
    return jsonify({'error': 'método no permitido'}), 405


@bp.route('/api/maestro-mps', methods=['GET','POST'])
def handle_maestro():
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        u, err, code = _require_planta_write()
        if err:
            return err, code
        d = request.json or {}
        # Validaciones zero-error · Sebastian 4-may-2026
        codigo = (d.get('codigo_mp') or '').strip().upper()
        if not codigo:
            return jsonify({'error': 'codigo_mp requerido'}), 400
        if len(codigo) > 50:
            return jsonify({'error': 'codigo_mp muy largo (max 50)'}), 400
        nombre_comercial = (d.get('nombre_comercial') or '').strip()
        nombre_inci = (d.get('nombre_inci') or '').strip()
        if not nombre_comercial and not nombre_inci:
            return jsonify({'error': 'Al menos un nombre (comercial o INCI) es requerido'}), 400
        try:
            stock_minimo = float(d.get('stock_minimo', 0) or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'stock_minimo invalido'}), 400
        if stock_minimo < 0:
            return jsonify({'error': 'stock_minimo no puede ser negativo'}), 400
        try:
            precio_referencia = float(d.get('precio_referencia', 0) or 0)
        except (TypeError, ValueError):
            return jsonify({'error': 'precio_referencia invalido'}), 400
        if precio_referencia < 0:
            return jsonify({'error': 'precio_referencia no puede ser negativo'}), 400
        # tipo_material: validar contra lista permitida
        tipo_material = d.get('tipo_material', 'MP')
        if tipo_material not in ('MP', 'Envase Primario', 'Envase Secundario', 'Empaque'):
            tipo_material = 'MP'

        # Idempotencia controlada: detectar si ya existe
        existente = c.execute(
            "SELECT codigo_mp, nombre_comercial, activo FROM maestro_mps WHERE codigo_mp=?",
            (codigo,)
        ).fetchone()
        if existente:
            # Si esta archivada (activo=0) y le piden crear igual, reactivar
            if not d.get('forzar_actualizar') and existente[2]:
                return jsonify({
                    'error': f'Ya existe MP con código {codigo}: "{existente[1]}". '
                             'Pasa forzar_actualizar=true para sobrescribir, o usa código diferente.',
                    'existente': {'codigo_mp': existente[0], 'nombre_comercial': existente[1]},
                }), 409
            # forzar o estaba archivada → UPDATE (reactiva)
            c.execute("""UPDATE maestro_mps
                         SET nombre_inci=?, nombre_comercial=?, tipo=?, proveedor=?,
                             stock_minimo=?, tipo_material=?, precio_referencia=?,
                             activo=1
                         WHERE codigo_mp=?""",
                      (nombre_inci, nombre_comercial, d.get('tipo',''),
                       d.get('proveedor',''), stock_minimo, tipo_material,
                       precio_referencia, codigo))
            accion = 'ACTUALIZAR_MP'
            mensaje = f'MP {codigo} actualizada' + (' y reactivada' if not existente[2] else '')
        else:
            # INSERT nuevo
            c.execute("""INSERT INTO maestro_mps
                         (codigo_mp, nombre_inci, nombre_comercial, tipo, proveedor,
                          stock_minimo, tipo_material, precio_referencia, activo)
                         VALUES (?,?,?,?,?,?,?,?,1)""",
                      (codigo, nombre_inci, nombre_comercial, d.get('tipo',''),
                       d.get('proveedor',''), stock_minimo, tipo_material,
                       precio_referencia))
            accion = 'CREAR_MP'
            mensaje = f'MP {codigo} creada'

        # audit_log obligatorio para cambios en maestro
        try:
            from audit_helpers import audit_log
            audit_log(c, usuario=u, accion=accion, tabla='maestro_mps',
                      registro_id=codigo,
                      despues={'codigo': codigo, 'nombre': nombre_comercial or nombre_inci,
                                'tipo_material': tipo_material,
                                'stock_minimo': stock_minimo,
                                'precio_referencia': precio_referencia,
                                'proveedor': d.get('proveedor', '')[:80]},
                      detalle=f"{accion} {codigo} · {nombre_comercial or nombre_inci}")
        except Exception:
            pass
        conn.commit()
        return jsonify({
            'ok': True, 'message': mensaje,
            'codigo_mp': codigo,
            'tipo_material': tipo_material,
            'creada': accion == 'CREAR_MP',
        }), 201
    # GET: filtro opcional por tipo_material via query param
    tipo_filter = (request.args.get('tipo_material') or '').strip()
    if tipo_filter and tipo_filter in ('MP', 'Envase Primario', 'Envase Secundario', 'Empaque'):
        c.execute("""SELECT codigo_mp,nombre_inci,nombre_comercial,tipo,proveedor,stock_minimo,
                            COALESCE(precio_referencia,0), COALESCE(tipo_material,'MP')
                     FROM maestro_mps WHERE activo=1 AND tipo_material=?
                     ORDER BY nombre_comercial COLLATE NOCASE""", (tipo_filter,))
    else:
        c.execute("""SELECT codigo_mp,nombre_inci,nombre_comercial,tipo,proveedor,stock_minimo,
                            COALESCE(precio_referencia,0), COALESCE(tipo_material,'MP')
                     FROM maestro_mps WHERE activo=1
                     ORDER BY nombre_comercial COLLATE NOCASE""")
    rows = c.fetchall()
    return jsonify({'mps': [
        {'codigo_mp':r[0], 'nombre_inci':r[1], 'nombre_comercial':r[2],
         'tipo':r[3], 'proveedor':r[4], 'stock_minimo':r[5],
         'precio_referencia':r[6], 'tipo_material':r[7]}
        for r in rows
    ]})

@bp.route('/api/maestro-mps/<codigo>')
def get_mp(codigo):
    # SEC-FIX · 21-may-2026 · auth obligatorio
    u, err, code_a = _require_session()
    if err: return err, code_a
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT codigo_mp,nombre_inci,nombre_comercial,tipo,proveedor,stock_minimo FROM maestro_mps WHERE codigo_mp=?", (codigo,))
    r = c.fetchone()
    return jsonify({'codigo_mp':r[0],'nombre_inci':r[1],'nombre_comercial':r[2],'tipo':r[3],'proveedor':r[4],'stock_minimo':r[5]}) if r else (jsonify({'error':'not found'}),404)

@bp.route('/api/maestro-mps/<codigo>/stock-minimo', methods=['PUT'])
def update_stock_minimo(codigo):
    """Actualiza el stock minimo de una MP."""
    # SEC-FIX · 21-may-2026 · auth + audit_log obligatorio
    # Antes: cualquier request PUT modificaba umbrales · sin rastro
    # Riesgo: manipulación silenciosa de flujo de compras
    u, err, code_a = _require_session()
    if err: return err, code_a
    if u not in (set(COMPRAS_USERS) | set(ADMIN_USERS)):
        return jsonify({'error': 'Solo Compras/Admin pueden editar mínimos'}), 403
    d = request.json or {}
    nuevo_min = d.get('stock_minimo')
    if nuevo_min is None:
        return jsonify({'error': 'stock_minimo requerido'}), 400
    conn = get_db(); c = conn.cursor()
    # Capturar valor anterior para audit
    prev_row = c.execute("SELECT stock_minimo FROM maestro_mps WHERE codigo_mp=?", (codigo,)).fetchone()
    if not prev_row:
        return jsonify({'error': 'MP no encontrada'}), 404
    prev_min = float(prev_row[0] or 0)
    c.execute("UPDATE maestro_mps SET stock_minimo=? WHERE codigo_mp=?", (float(nuevo_min), codigo))
    if c.rowcount == 0:
        return jsonify({'error': 'MP no encontrada'}), 404
    try:
        audit_log(c, usuario=u, accion='UPDATE_STOCK_MINIMO',
                  tabla='maestro_mps', registro_id=codigo,
                  antes={'stock_minimo': prev_min},
                  despues={'stock_minimo': float(nuevo_min)})
    except Exception as _e:
        __import__('logging').getLogger('inventario').warning('audit update_stock_minimo: %s', _e)
    conn.commit()
    return jsonify({'message': f'Stock minimo de {codigo} actualizado a {nuevo_min}'})

@bp.route('/api/maestro-mps/<codigo>/proveedor', methods=['PUT'])
def update_mp_proveedor(codigo):
    """Actualiza el proveedor asignado a una MP en maestro_mps.

    Caso de uso original: edicion del catalogo desde dashboard.
    Caso de uso 2026-04-27 (CEO): tambien usado per-item desde modal
    Ver & Gestionar de una solicitud cuando alguna MP esta trocada
    (asignada al proveedor incorrecto). Por eso ahora deja audit_log.

    Body:
      proveedor: str

    Si la MP no existe en maestro_mps, la crea (legacy behavior preservado
    para soportar entradas que solo estaban en movimientos historicos).
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    proveedor = (d.get('proveedor') or '').strip()
    conn = get_db()
    c = conn.cursor()

    # Snapshot del valor anterior antes del update
    anterior_row = c.execute(
        "SELECT proveedor FROM maestro_mps WHERE codigo_mp=?", (codigo,)
    ).fetchone()
    anterior = (anterior_row[0] or '') if anterior_row else None

    # 1. Actualizar maestro_mps
    c.execute("UPDATE maestro_mps SET proveedor=? WHERE codigo_mp=?", (proveedor, codigo))
    updated = c.rowcount

    if updated == 0:
        # MP no existe en maestro_mps — crearla con info de movimientos
        mov = c.execute(
            "SELECT material_nombre FROM movimientos WHERE material_id=? LIMIT 1", (codigo,)
        ).fetchone()
        nombre = mov[0] if mov else codigo
        c.execute("""
            INSERT INTO maestro_mps (codigo_mp, nombre_comercial, nombre_inci, tipo, proveedor, stock_minimo, activo)
            VALUES (?, ?, '', 'MP', ?, 0, 1)
            ON CONFLICT(codigo_mp) DO UPDATE SET proveedor=excluded.proveedor
        """, (codigo, nombre, proveedor))

    # 2. Upsert en directorio de proveedores (tabla proveedores) si tiene nombre
    if proveedor:
        from datetime import datetime as _dt
        # Comparacion case-insensitive · Catalina 4-may-2026
        exists = c.execute(
            "SELECT nombre FROM proveedores WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?))",
            (proveedor,)
        ).fetchone()
        if not exists:
            try:
                c.execute("""
                    INSERT INTO proveedores
                    (nombre, contacto, email, telefono, categoria, condiciones_pago,
                     nit, direccion, num_cuenta, tipo_cuenta, banco, concepto_compra, fecha_creacion)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (proveedor, '', '', '', 'mp', '30 dias',
                      '', '', '', '', '', 'Materias Primas', _dt.now().isoformat()))
                try:
                    from audit_helpers import audit_log as _al
                    _al(c, usuario=session.get('compras_user', 'sistema'),
                        accion='CREAR_PROVEEDOR', tabla='proveedores',
                        registro_id=c.lastrowid,
                        despues={'nombre': proveedor[:200],
                                  'origen': 'auto_setear_proveedor_mp',
                                  'codigo_mp': str(codigo)[:50]},
                        detalle=f"Auto-creado al asignar proveedor a MP {codigo}")
                except Exception:
                    pass
            except Exception:
                pass  # Si ya existe por nombre con diferente case, ignorar

    # Audit log si hubo cambio real
    if anterior is None or anterior != proveedor:
        try:
            import json as _json
            c.execute("""INSERT INTO audit_log
                         (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                         VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                      (user, 'EDITAR_PROVEEDOR_MP', 'maestro_mps', codigo,
                       _json.dumps({
                           'codigo_mp': codigo,
                           'proveedor_anterior': anterior or '',
                           'proveedor_nuevo': proveedor,
                           'mp_creada': anterior is None,
                       }, ensure_ascii=False),
                       request.remote_addr))
        except sqlite3.OperationalError:
            pass

    conn.commit()
    return jsonify({
        'ok': True, 'codigo_mp': codigo, 'proveedor': proveedor,
        'proveedor_anterior': anterior or '',
    })

@bp.route('/api/maestro-mps/<codigo>/mee-stock-minimo', methods=['PUT'])
def update_mee_stock_minimo(codigo):
    """Actualiza el stock minimo de un MEE."""
    d = request.json or {}
    nuevo_min = d.get('stock_minimo')
    if nuevo_min is None:
        return jsonify({'error': 'stock_minimo requerido'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE maestro_mee SET stock_minimo=? WHERE codigo=?", (float(nuevo_min), codigo))
    if c.rowcount == 0:
        return jsonify({'error': 'MEE no encontrado'}), 404
    conn.commit()
    return jsonify({'message': f'Stock minimo de {codigo} actualizado a {nuevo_min}'})

@bp.route('/api/consumo-manual', methods=['POST'])
def consumo_manual():
    """Registra consumo manual de una MP (ajuste por uso).

    Sebastian 5-may-2026 (audit zero-error Bodega MP): ANTES no tenia
    proteccion de permisos NI audit_log · cualquier request anonimo
    podia descontar stock sin rastro. Brecha grave de trazabilidad
    INVIMA (Resolucion 2214/2021 art. 10).

    FIX: requerir permiso planta_write + audit_log obligatorio +
    validacion del codigo MP existe en catalogo.
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code

    d = request.json or {}
    codigo = (d.get('codigo_mp') or '').upper().strip()
    cantidad = float(d.get('cantidad') or 0)
    lote = (d.get('lote') or '').strip()
    obs = (d.get('observaciones') or 'Consumo manual').strip()
    operador = d.get('operador', '').strip() or u
    if not codigo or cantidad <= 0:
        return jsonify({'error': 'Codigo y cantidad positiva requeridos'}), 400

    conn = get_db(); c = conn.cursor()
    c.execute("SELECT nombre_comercial FROM maestro_mps WHERE codigo_mp=?", (codigo,))
    mp = c.fetchone()
    if not mp:
        return jsonify({'error': f'MP {codigo} no existe en catalogo'}), 404
    nombre = mp[0] or codigo

    # Calcular stock disponible ANTES del descuento (para audit + warning).
    # B-1 (Sebastian 12-jun): excluir lotes retenidos (cuarentena/rechazado/vencido/
    # agotado/bloqueado) · antes sumaba TODO -> permitia descontar stock que Calidad
    # tiene retenido y dejar el saldo producible negativo. Alineado con /api/lotes.
    stock_antes_row = c.execute(
        "SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0) "
        "FROM movimientos WHERE material_id=? "
        "AND (estado_lote IS NULL OR UPPER(COALESCE(estado_lote,'')) NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO'))",
        (codigo,),
    ).fetchone()
    stock_antes = float(stock_antes_row[0] or 0)
    if cantidad > stock_antes:
        # SEC-FIX · 21-may-2026 · bloquear stock negativo · flag forzar para excepciones
        # Antes: solo loguea warning · permite stock fantasma · drift FEFO/ABC/costing
        if not d.get('forzar_sobreconsumo'):
            return jsonify({
                'error': f'Stock insuficiente · disponible {stock_antes:.2f}g · solicitado {cantidad:.2f}g',
                'codigo': 'STOCK_INSUFICIENTE',
                'stock_disponible': stock_antes,
                'hint': 'Si es ajuste correctivo deliberado, enviar forzar_sobreconsumo=true (admin)',
            }), 422
        # Solo admins pueden forzar
        if u not in ADMIN_USERS:
            return jsonify({
                'error': 'Solo admin puede forzar consumo > stock',
                'codigo': 'FORZAR_SOLO_ADMIN',
            }), 403
        __import__('logging').getLogger('inventario').warning(
            "consumo_manual FORZADO descontando MAS de lo disponible · codigo=%s "
            "cantidad=%s stock=%s usuario=%s",
            codigo, cantidad, stock_antes, u,
        )

    c.execute("""INSERT INTO movimientos
                 (material_id,material_nombre,cantidad,tipo,fecha,observaciones,lote,operador)
                 VALUES (?,?,?,'Salida',datetime('now', '-5 hours'),?,?,?)""",
              (codigo, nombre, cantidad, obs, lote, operador))
    mov_id = c.lastrowid

    try:
        audit_log(
            c, usuario=u, accion='CONSUMO_MANUAL',
            tabla='movimientos', registro_id=mov_id,
            despues={
                'codigo_mp': codigo,
                'nombre': nombre[:200],
                'cantidad_g': cantidad,
                'lote': lote or '',
                'observaciones': obs[:200],
                'operador': operador[:100],
                'stock_antes_g': round(stock_antes, 2),
                'stock_despues_g': round(stock_antes - cantidad, 2),
                'sobreconsumo': cantidad > stock_antes,
            },
            detalle=(f"Consumo manual: {codigo} · {cantidad:.0f}g" +
                      (f" · lote {lote}" if lote else "") +
                      (" · ⚠ EXCEDE stock disponible"
                       if cantidad > stock_antes else "")),
        )
    except Exception as _ae:
        __import__('logging').getLogger('inventario').warning(
            "audit_log CONSUMO_MANUAL fallo · mov_id=%s err=%s", mov_id, _ae,
        )

    conn.commit()
    return jsonify({
        'message': f'Consumo de {cantidad}g registrado para {nombre}',
        'mov_id': mov_id,
        'codigo_mp': codigo,
        'cantidad_g': cantidad,
        'stock_despues_g': round(stock_antes - cantidad, 2),
    }), 201

@bp.route('/api/maestro-mps/<codigo>/archivar', methods=['PUT'])
def archivar_mp(codigo):
    """Archiva una MP (la marca como inactiva sin borrarla).

    Solo Calidad/Admin · archivar afecta el catálogo regulatorio (la MP
    queda invisible para nuevas producciones). Auditado.
    """
    u, err, code = _require_qc()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    antes_row = c.execute(
        "SELECT codigo_mp, nombre_comercial, activo FROM maestro_mps WHERE codigo_mp=?",
        (codigo,)).fetchone()
    if not antes_row:
        return jsonify({'error': 'MP no encontrada'}), 404
    antes = dict(antes_row)
    c.execute("UPDATE maestro_mps SET activo=0 WHERE codigo_mp=?", (codigo,))
    if c.rowcount == 0:
        return jsonify({'error': 'MP no encontrada'}), 404
    audit_log(c, usuario=u, accion='ARCHIVAR_MP', tabla='maestro_mps',
              registro_id=codigo, antes=antes,
              despues={'activo': 0},
              detalle=f"Archivó MP {codigo} ({antes.get('nombre_comercial','')})")
    conn.commit()
    return jsonify({'message': f'MP {codigo} archivada exitosamente'})

@bp.route('/api/recepcion', methods=['POST'])
def registrar_recepcion():
    """Sprint Recepciones PRO · 20-may-2026 · validaciones reforzadas:
    - #1 duplicado lote+material_id (ya existía 10min, ampliado)
    - #2 cantidad ≤ pendiente de la OC (warning hard)
    - #3 cuarentena → push_notif a Calidad
    - #4 factura obligatoria si vinculó OC
    - #12 alerta precio cambió >30% vs último ingreso
    """
    u, err, code = _require_planta_write()
    if err:
        return err, code
    d = request.json; codigo = (d.get('codigo_mp') or '').upper().strip()
    if not codigo: return jsonify({'error': 'Codigo MP requerido'}), 400
    cantidad_recibida = float(d.get('cantidad') or 0)
    if cantidad_recibida <= 0:
        return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400
    conn = get_db(); c = conn.cursor()

    # Sprint Recepciones PRO fix #4: factura obligatoria si hay OC
    numero_factura_pre = (d.get('numero_factura') or '').strip()
    numero_oc_pre = (d.get('numero_oc') or '').strip()
    if numero_oc_pre and not numero_factura_pre and not d.get('factura_omitida_explicita'):
        return jsonify({
            'error': (f'Factura obligatoria · vinculaste OC {numero_oc_pre} '
                      'pero no llenaste el campo "N° Factura/Remisión". '
                      'Audit contable lo requiere.'),
            'factura_obligatoria': True,
        }), 400

    # Sprint Recepciones PRO fix #2: cantidad ≤ pendiente de la OC
    if numero_oc_pre and not d.get('forzar_excedente'):
        try:
            row_pend = c.execute(
                """SELECT COALESCE(SUM(cantidad_g - COALESCE(cantidad_recibida_g,0)), 0)
                   FROM ordenes_compra_items
                   WHERE numero_oc = ? AND codigo_mp = ?""",
                (numero_oc_pre, codigo),
            ).fetchone()
            pendiente_g = float(row_pend[0] or 0)
            if pendiente_g > 0 and cantidad_recibida > pendiente_g * 1.05:
                # Tolerancia 5% (frascos pueden venir con leve sobrante)
                return jsonify({
                    'error': (
                        f'Cantidad excede lo pendiente de OC {numero_oc_pre} · '
                        f'recibís {cantidad_recibida:.0f}g pero quedan '
                        f'{pendiente_g:.0f}g por recibir. Si recibiste de más, '
                        'reintentá con forzar_excedente=true.'),
                    'cantidad_excede_oc': True,
                    'pendiente_oc_g': pendiente_g,
                }), 409
        except Exception as _e_sobre:
            # REC-02 (12-jun · M4): no tragar mudo · si la query de pendiente-OC
            # falla, dejamos rastro (antes perdiamos la validacion de sobre-recepcion
            # en silencio · p.ej. drift de columnas en PG). No bloqueante.
            __import__('logging').getLogger('inventario').warning(
                "registrar_recepcion sobre-recepcion check fallo (no bloquea): %s", _e_sobre)

    # Sprint Recepciones PRO fix #12: alerta si precio cambió >30%
    precio_pre = float(d.get('precio_kg') or 0)
    alerta_precio = None
    if precio_pre > 0:
        try:
            prev_row = c.execute(
                """SELECT precio_kg FROM precios_mp_historico
                   WHERE codigo_mp = ? AND precio_kg > 0
                   ORDER BY fecha DESC LIMIT 1""",
                (codigo,),
            ).fetchone()
            if prev_row and prev_row[0]:
                prev_precio = float(prev_row[0])
                if prev_precio > 0:
                    delta_pct = abs(precio_pre - prev_precio) / prev_precio * 100
                    if delta_pct >= 30:
                        direccion = 'subió' if precio_pre > prev_precio else 'bajó'
                        alerta_precio = (
                            f'⚠ Precio {direccion} {delta_pct:.0f}% vs último '
                            f'ingreso (${prev_precio:,.0f}/kg → ${precio_pre:,.0f}/kg). '
                            'Verificá que sea correcto.'
                        )
        except Exception:
            pass
    c.execute("SELECT nombre_inci,nombre_comercial,tipo,proveedor FROM maestro_mps WHERE codigo_mp=?", (codigo,))
    mp = c.fetchone()
    nombre_comercial = d.get('nombre_comercial','') or (mp[1] if mp else codigo)
    # El kardex se rotula por INCI (Sebastian 12-jun · en recepcion NO se usa el
    # comercial, que varia por proveedor y confunde); cae al codigo si no hay INCI.
    # El comercial NO se borra: se sigue guardando en maestro_mps.nombre_comercial.
    _inci_rec = (mp[0] if mp else d.get('nombre_inci','')) or ''
    nombre = _inci_rec.strip() or codigo
    proveedor = d.get('proveedor','') or (mp[3] if mp else '')
    precio_kg = float(d.get('precio_kg') or 0)
    numero_factura = (d.get('numero_factura') or '').strip()
    numero_oc = (d.get('numero_oc') or '').strip()
    # FIX 13-jun (audit compras · INVIMA cuarentena-first): default = CUARENTENA.
    # La recepción por OC entra en cuarentena; el ingreso manual es consistente
    # (MP recibida pasa por QC antes de usarse). El operario puede destildar la
    # casilla para stock ya aprobado (ajustes/correcciones).
    # Sebastián 16-jun · si RECEPCION_AUTO_VIGENTE está encendido, el default de la
    # casilla pasa a NO-cuarentena → carga automática como VIGENTE (sin Calidad).
    from database import recepcion_auto_vigente as _rav
    _cuar_default = (not _rav(c))
    cuarentena = bool(d.get('cuarentena', _cuar_default))
    estado_lote = 'CUARENTENA' if cuarentena else 'VIGENTE'
    # Si la MP es nueva y viene con datos, crearla en el catalogo
    if not mp and (d.get('nombre_inci') or d.get('nombre_comercial')):
        c.execute("INSERT OR IGNORE INTO maestro_mps (codigo_mp, nombre_inci, nombre_comercial, tipo, proveedor, stock_minimo) VALUES (?,?,?,?,?,?)",
                  (codigo, d.get('nombre_inci',''), nombre_comercial, d.get('tipo',''), proveedor, d.get('stock_minimo',0)))
        conn.commit()
    # Actualizar precio_referencia en maestro_mps si viene precio.
    # Solo ignoramos si la columna 'ultima_act_precio' no existe (versión vieja).
    if precio_kg > 0:
        try:
            c.execute("UPDATE maestro_mps SET precio_referencia=?, ultima_act_precio=datetime('now', '-5 hours') WHERE codigo_mp=?", (precio_kg, codigo))
        except sqlite3.OperationalError as _e:
            if 'no such column' not in str(_e).lower():
                __import__('logging').getLogger('inventario').error(
                    "UPDATE precio_referencia falló para %s: %s", codigo, _e
                )
    lote = (d.get('lote') or '').strip()
    if not lote or lote.upper()=='AUTO':
        from datetime import date; lote = f"ESP{date.today().strftime('%y%m%d')}{codigo[-3:]}"
    # REC-01 (12-jun · INVIMA): bloquear MP YA VENCIDA en el ingreso manual
    # (puerta lateral · recibir_oc ya lo bloquea). Antes registrar_recepcion
    # aceptaba fecha_vencimiento < hoy y la metia como VIGENTE -> material
    # vencido usable en FEFO sin alerta. Override admin: forzar_vencido=true.
    _fv = (d.get('fecha_vencimiento') or '').strip()
    if _fv and len(_fv) >= 10 and not d.get('forzar_vencido'):
        try:
            from datetime import date as _dvenc
            if _dvenc.fromisoformat(_fv[:10]) < _dvenc.today():
                return jsonify({
                    'error': 'fecha_vencimiento ya pasada (' + _fv[:10] + ') · MP vencida no debe ingresar como disponible',
                    'vencimiento_pasado': True,
                    'hint': 'Si es un ingreso historico/excepcional, reenvia con forzar_vencido=true (admin).',
                }), 409
        except (ValueError, TypeError):
            pass  # fecha mal formada · no bloquear por eso (la validan otros)
    # Sebastián 15-may-2026: guard de idempotencia. Tras el incidente de
    # corrupción de BD, si la analista vio "error interno" y reintentó un
    # ingreso que en realidad sí se había guardado, se duplicaba el
    # movimiento de Entrada (stock inflado al doble). Ahora: si ya existe
    # una Entrada idéntica (material+lote+cantidad) en los últimos 10 min,
    # se rechaza con 409 salvo que se pase forzar_duplicado=true.
    if not d.get('forzar_duplicado'):
        # FIX 30-may-2026 · audit Planta/Recepciones · ANTES usaba
        # julianday('now')-julianday(fecha), que es SQLite-only · en PostgreSQL
        # (producción) la función no existe → la recepción tiraba 500 en cada
        # ingreso no-forzado. Ahora: cutoff de 10 min calculado en Python y
        # comparación de strings ISO (fecha se guarda como datetime.now().isoformat()).
        from datetime import timedelta as _tddup
        _cut_dup = (datetime.now() - _tddup(minutes=10)).isoformat()
        dup = c.execute(
            """SELECT id, fecha FROM movimientos
               WHERE material_id=? AND lote=? AND cantidad=? AND tipo='Entrada'
                 AND fecha >= ?
               ORDER BY id DESC LIMIT 1""",
            (codigo, lote, cantidad_recibida, _cut_dup),
        ).fetchone()
        if dup:
            return jsonify({
                'error': (f'Ya hay un ingreso idéntico de {codigo} · lote {lote} · '
                          f'{cantidad_recibida:.0f}g registrado hace menos de 10 min '
                          f'(movimiento #{dup[0]}). Si de verdad querés registrarlo '
                          f'otra vez, reintentá marcando "forzar duplicado".'),
                'posible_duplicado': True,
                'movimiento_existente_id': dup[0],
            }), 409
    c.execute("""INSERT INTO movimientos
                 (material_id,material_nombre,cantidad,tipo,fecha,observaciones,
                  lote,fecha_vencimiento,estanteria,posicion,proveedor,estado_lote,operador,
                  precio_kg,numero_factura,numero_oc)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
              (codigo,nombre,cantidad_recibida,'Entrada',datetime.now().isoformat(),
               d.get('observaciones','Ingreso MP'),lote,d.get('fecha_vencimiento',''),
               d.get('estanteria',''),d.get('posicion',''),proveedor,estado_lote,
               d.get('operador',''),precio_kg,numero_factura,numero_oc))
    mov_id = c.lastrowid
    # Log precio historico — solo ignorar si tabla no existe (legacy).
    if precio_kg > 0:
        try:
            c.execute("INSERT OR IGNORE INTO precios_mp_historico (codigo_mp,precio_kg,numero_factura,proveedor,fecha) VALUES (?,?,?,?,datetime('now', '-5 hours'))",
                      (codigo, precio_kg, numero_factura, proveedor))
        except sqlite3.OperationalError as _e:
            if 'no such table' not in str(_e).lower():
                __import__('logging').getLogger('inventario').error(
                    "INSERT precios_mp_historico falló: %s", _e
                )
    # Cerrar OC si se referencia una
    oc_warning = None
    if numero_oc:
        try:
            c.execute("UPDATE ordenes_compra_items SET cantidad_recibida_g=cantidad_recibida_g+?,lote_asignado=? WHERE numero_oc=? AND codigo_mp=?",
                      (cantidad_recibida, lote, numero_oc, codigo))
            # verificar si todos los items de la OC estan recibidos
            c.execute("SELECT COUNT(*) FROM ordenes_compra_items WHERE numero_oc=? AND (cantidad_g - cantidad_recibida_g) > 1", (numero_oc,))
            pendientes = c.fetchone()[0]
            if pendientes == 0:
                # FIX 1-jun-2026 audit · 'Recibida' (no 'RECIBIDA') · el resto del
                # sistema usa mixed-case · una OC en 'RECIBIDA' desaparecía de listas
                # de recibidas / cola de pago / discrepancias / aprendizaje de lead-time.
                c.execute("UPDATE ordenes_compra SET estado='Recibida',fecha_recepcion=datetime('now', '-5 hours'),recibido_por=? WHERE numero_oc=?",
                          (d.get('operador',''), numero_oc))
        except Exception as oc_err:
            # Log but don't fail the reception — OC can be reconciled manually
            print(f'[WARN] OC update failed for {numero_oc}: {oc_err}', flush=True)
            oc_warning = f'OC {numero_oc} no pudo actualizarse automaticamente — verificar manualmente'
    # Sebastian 5-may-2026 (audit zero-error Bodega MP): ANTES /api/recepcion
    # NO generaba audit_log a pesar de ser un evento INVIMA-critico (entrada
    # fisica de MP, trazabilidad GMP/BPM Resolucion 2214/2021 art. 10).
    # FIX: registrar quien ingreso que cantidad de que MP, en que lote, con
    # que proveedor + OC. Permite rastreo de re-llamados y auditoria QC.
    try:
        audit_log(
            c, usuario=u, accion='INGRESAR_MP',
            tabla='movimientos', registro_id=mov_id,
            despues={
                'codigo_mp': codigo,
                'nombre': nombre[:200],
                'cantidad_g': cantidad_recibida,
                'lote': lote,
                'proveedor': (proveedor or '')[:200],
                'numero_oc': numero_oc or '',
                'numero_factura': numero_factura or '',
                'fecha_vencimiento': d.get('fecha_vencimiento', '') or '',
                'cuarentena': cuarentena,
                'estado_lote': estado_lote,
                'precio_kg': precio_kg,
            },
            detalle=(f"Ingreso MP: {codigo} · {cantidad_recibida:.0f}g · "
                      f"lote {lote}" +
                      (f" · OC {numero_oc}" if numero_oc else "") +
                      (" · CUARENTENA" if cuarentena else "")),
        )
    except Exception as _ae:
        __import__('logging').getLogger('inventario').warning(
            "audit_log INGRESAR_MP fallo · mov_id=%s err=%s", mov_id, _ae,
        )
    conn.commit()

    # Sprint Recepciones PRO fix #3: push_notif a Calidad si cuarentena
    if cuarentena:
        try:
            from blueprints.notif import push_notif as _push_notif
            from config import CALIDAD_USERS
            for dest in (CALIDAD_USERS or set()):
                _push_notif(
                    destinatario=dest,
                    tipo='cuarentena_nueva',
                    titulo=f'🔒 MP nueva en cuarentena · {codigo}',
                    body=(f'{nombre} · lote {lote} · {cantidad_recibida:.0f}g · '
                          f'{proveedor or "sin proveedor"}'),
                    link='/dashboard#cuarentena',
                    remitente=u,
                    importante=True,
                )
        except Exception as _ne:
            __import__('logging').getLogger('inventario').warning(
                'push_notif cuarentena fallo: %s', _ne)

    msg = f'{nombre} ingresada. Lote: {lote}'
    if cuarentena: msg += ' — En CUARENTENA (pendiente aprobacion QC)'
    if numero_oc and not oc_warning: msg += f' | OC {numero_oc} actualizada'
    if oc_warning: msg += f' | ⚠ {oc_warning}'
    return jsonify({
        'message': msg, 'lote': lote, 'codigo': codigo, 'nombre': nombre,
        'cantidad': cantidad_recibida, 'cuarentena': cuarentena,
        'oc_warning': oc_warning,
        'mov_id': mov_id,
        'alerta_precio': alerta_precio,
    }), 201

@bp.route('/api/recepcion/recientes', methods=['GET'])
def recepcion_recientes():
    """Sprint Recepciones PRO · 20-may-2026 · fix #7 + #13.

    Endpoint dedicado para "Últimas entradas" con paginación y búsqueda
    server-side. Antes el frontend bajaba TODOS los movimientos y filtraba
    en JS · perf horrible.

    Query params:
      limit (default 25, max 200)
      offset (default 0)
      q: busca en material_nombre, material_id, lote, proveedor (case-insensitive)
    """
    u, err, code = _require_session()
    if err:
        return err, code
    try:
        limit = max(1, min(int(request.args.get('limit', 25)), 200))
    except (ValueError, TypeError):
        limit = 25
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    q = (request.args.get('q') or '').strip().lower()
    params = []
    where = ["m.tipo IN ('Entrada','entrada','ENTRADA')"]
    if q:
        # Escape LIKE wildcards (% y _) para evitar match fantasma
        q_safe = q.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
        like = f'%{q_safe}%'
        where.append(
            "(LOWER(COALESCE(m.material_nombre,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(m.material_id,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(m.lote,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(m.proveedor,'')) LIKE ? ESCAPE '\\')",
        )
        params.extend([like, like, like, like])
    where_sql = ' AND '.join(where)
    conn = get_db(); c = conn.cursor()
    # Total count para paginación
    try:
        total = c.execute(
            f"SELECT COUNT(*) FROM movimientos m WHERE {where_sql}", params,
        ).fetchone()[0]
    except Exception:
        total = 0
    # Datos con OC vinculada (LEFT JOIN para no perder rows sin OC)
    try:
        rows = c.execute(
            f"""SELECT m.id, m.material_id, m.material_nombre, m.lote,
                       m.cantidad, COALESCE(m.proveedor,'') as proveedor,
                       COALESCE(m.fecha_vencimiento,'') as venc,
                       m.fecha, COALESCE(m.estado_lote,'') as estado_lote,
                       COALESCE(m.numero_oc,'') as numero_oc,
                       COALESCE(m.numero_factura,'') as numero_factura,
                       COALESCE(m.precio_kg, 0) as precio_kg,
                       COALESCE(m.operador,'') as operador,
                       COALESCE(mp.nombre_inci,'') as nombre_inci
                FROM movimientos m
                LEFT JOIN maestro_mps mp ON m.material_id = mp.codigo_mp
                WHERE {where_sql}
                ORDER BY m.fecha DESC, m.id DESC
                LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
    except Exception as _e:
        rows = []
    items = [{
        'id': r[0], 'material_id': r[1] or '',
        'material_nombre': r[2] or '', 'lote': r[3] or '',
        'cantidad_g': float(r[4] or 0), 'proveedor': r[5] or '',
        'fecha_vencimiento': r[6] or '', 'fecha': r[7] or '',
        'estado_lote': r[8] or '', 'numero_oc': r[9] or '',
        'numero_factura': r[10] or '', 'precio_kg': float(r[11] or 0),
        'operador': r[12] or '', 'nombre_inci': r[13] or '',
    } for r in rows]
    return jsonify({
        'items': items, 'total': total, 'limit': limit, 'offset': offset,
        'has_more': (offset + len(items)) < total,
    })


@bp.route('/api/recepcion/<int:mov_id>/anular', methods=['POST'])
def anular_recepcion(mov_id):
    """Sprint Recepciones PRO · 20-may-2026 · fix #8 · admin.

    Anula una recepción ingresada por error. NO borra el movimiento original
    (audit reversible) · crea movimiento Salida inverso con observaciones
    "ANULACIÓN: <motivo>". Si la recepción venía de OC, descuenta también el
    cantidad_recibida_g de ordenes_compra_items.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    if u not in ADMIN_USERS:
        return jsonify({'error': 'solo admin'}), 403
    body = request.get_json(silent=True) or {}
    motivo = (body.get('motivo') or '').strip()
    if len(motivo) < 10:
        return jsonify({'error': 'motivo requerido (≥10 chars)'}), 400
    conn = get_db(); c = conn.cursor()
    # Buscar movimiento original
    row = c.execute(
        """SELECT material_id, material_nombre, cantidad, tipo, lote,
                  proveedor, COALESCE(numero_oc,''), COALESCE(numero_factura,''),
                  COALESCE(estado_lote,'')
           FROM movimientos WHERE id = ?""",
        (mov_id,),
    ).fetchone()
    if not row:
        return jsonify({'error': f'movimiento {mov_id} no existe'}), 404
    if (row[3] or '').lower() != 'entrada':
        return jsonify({'error': 'solo se anula movimientos tipo Entrada'}), 400
    # Buscar si ya hay una anulación previa (idempotencia)
    prev = c.execute(
        """SELECT id FROM movimientos
           WHERE tipo='Salida' AND material_id=? AND lote=? AND cantidad=?
             AND COALESCE(observaciones,'') LIKE ?
           LIMIT 1""",
        (row[0], row[4], row[2], f'%ANULACI%mov#{mov_id}%'),
    ).fetchone()
    if prev:
        return jsonify({
            'error': f'ya hay anulación previa (mov #{prev[0]})',
            'anulacion_existente': prev[0],
        }), 409
    # FIX 13-jun (audit recepción · BUG-3): solo anular si la cantidad recibida SIGUE
    # disponible en el lote (no consumida). Stock RAW del lote (todas las filas, SIN
    # excluir estados · la cuarentena cuenta acá porque es lo recibido físico que vamos
    # a revertir). Evita dejar el stock NEGATIVO al anular un lote ya consumido.
    # ⚠ LIMITACIÓN CONOCIDA (P2): el RAW agrega por (material_id, lote); si el MISMO nº
    # de lote se reusa en 2 entregas distintas del mismo material y la 1ª ya se consumió,
    # el guard no distingue cuál recepción concreta se anula (el kardex FEFO no liga
    # Salida↔Entrada). Mitigado porque reusar nº de lote entre entregas es anti-patrón.
    _raw = c.execute(
        "SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad "
        "WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END),0) "
        "FROM movimientos WHERE material_id=? AND lote=?", (row[0], row[4]),
    ).fetchone()[0] or 0
    if float(_raw) + 0.01 < float(row[2] or 0):
        return jsonify({
            'error': (f'No se puede anular: el lote ya fue consumido/anulado · '
                      f'disponible {float(_raw):.0f}g < a anular {float(row[2] or 0):.0f}g'),
            'codigo': 'LOTE_YA_MOVIDO',
        }), 409
    # FIX 13-jun (P1 · revisión adversarial · M27): CLAIM atómico de la anulación sobre
    # la fila Entrada original, gate por rowcount==1. El check `prev` y el guard RAW de
    # arriba son check-then-act: en PostgreSQL (3 workers, READ COMMITTED) dos anulaciones
    # concurrentes del mismo mov_id pasaban AMBAS → doble Salida → stock NEGATIVO. Este
    # UPDATE condicional (observaciones NOT LIKE marcador) toma el row-lock: solo un worker
    # marca (rowcount==1); el 2º, tras el commit del 1º, ve rowcount==0 → 409. Idempotencia
    # real + anti-doble-anulación. El marcador queda en la Entrada como rastro permanente.
    _marca = f' ::ANULADA-mov#{mov_id}::'
    c.execute(
        "UPDATE movimientos SET observaciones = COALESCE(observaciones,'') || ? "
        "WHERE id=? AND COALESCE(observaciones,'') NOT LIKE ?",
        (_marca, mov_id, f'%::ANULADA-mov#{mov_id}::%'),
    )
    if c.rowcount != 1:
        conn.rollback()
        return jsonify({
            'error': 'esta recepción ya fue anulada (o se está anulando en paralelo)',
            'codigo': 'ANULACION_YA_RECLAMADA',
        }), 409
    # Insertar movimiento Salida inverso. FIX 13-jun (BUG-1/2): la Salida ESPEJA el
    # estado_lote original (no 'ANULADO') → anulación net-zero EXACTO en toda vista:
    # CUARENTENA→CUARENTENA (ambas excluidas) o VIGENTE→VIGENTE (ambas cuentan).
    # Antes 'ANULADO' restaba en el canónico (negativo) y no en auditar-minimos
    # (stock fantasma) → divergencia. La marca de anulación va en observaciones.
    obs_anul = (
        f'ANULACIÓN mov#{mov_id} · motivo: {motivo[:300]} · '
        f'por {u} {datetime.now().isoformat()}'
    )
    _estado_anul = (row[8] or 'VIGENTE')
    c.execute(
        """INSERT INTO movimientos
             (material_id, material_nombre, cantidad, tipo, fecha,
              observaciones, lote, proveedor, operador, estado_lote)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (row[0], row[1], row[2], 'Salida',
         datetime.now().isoformat(), obs_anul, row[4], row[5], u,
         _estado_anul),
    )
    mov_anul_id = c.lastrowid
    # Si tenía OC, descontar de cantidad_recibida_g
    numero_oc = row[6]
    if numero_oc:
        try:
            c.execute(
                """UPDATE ordenes_compra_items
                   SET cantidad_recibida_g = MAX(0, cantidad_recibida_g - ?)
                   WHERE numero_oc = ? AND codigo_mp = ?""",
                (row[2], numero_oc, row[0]),
            )
        except Exception:
            pass
    # Audit
    try:
        audit_log(
            c, usuario=u, accion='ANULAR_RECEPCION_MP',
            tabla='movimientos', registro_id=mov_id,
            antes={'tipo': 'Entrada', 'cantidad': row[2],
                   'lote': row[4], 'material_id': row[0]},
            despues={'mov_anulacion_id': mov_anul_id,
                     'motivo': motivo[:300],
                     'numero_oc': numero_oc,
                     'numero_factura': row[7]},
            detalle=f'Anulación recepción mov#{mov_id}: {motivo[:300]}',
        )
    except Exception as _ae:
        __import__('logging').getLogger('inventario').warning(
            'audit_log ANULAR_RECEPCION_MP fallo: %s', _ae)
    conn.commit()
    return jsonify({
        'ok': True,
        'mensaje': f'Recepción mov#{mov_id} anulada · mov anulación #{mov_anul_id}',
        'mov_anulacion_id': mov_anul_id,
    })


@bp.route('/api/recepcion/<codigo_mp>/precio-historico', methods=['GET'])
def precio_historico_mp(codigo_mp):
    """Sprint Recepciones PRO · 20-may-2026 · auxiliar para alerta precio.

    Devuelve últimos 10 precios de un codigo_mp · para que el frontend
    pueda mostrar tendencia o pre-validar antes de submit.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    try:
        rows = c.execute(
            """SELECT fecha, precio_kg, proveedor, numero_factura
               FROM precios_mp_historico
               WHERE codigo_mp = ? AND precio_kg > 0
               ORDER BY fecha DESC LIMIT 10""",
            (codigo_mp.upper(),),
        ).fetchall()
    except Exception:
        rows = []
    items = [{
        'fecha': r[0] or '', 'precio_kg': float(r[1] or 0),
        'proveedor': r[2] or '', 'numero_factura': r[3] or '',
    } for r in rows]
    return jsonify({'historial': items, 'codigo_mp': codigo_mp.upper()})


@bp.route('/api/inventario/diagnostico-post-incidente', methods=['GET'])
def diagnostico_post_incidente():
    """Detecta inconsistencias que pudo dejar el incidente de corrupción
    de BD del 15-may-2026. Read-only · no modifica nada.

    Sebastián 15-may-2026: "me preocupa que planta quedara con algo roto,
    revisemos antes de seguir". Chequea 3 cosas:
      1. Movimientos de Entrada duplicados (reintento tras error 500)
      2. Producciones iniciadas que quedaron sin terminar
      3. Lotes con stock negativo (imposible · señal de drift)
    """
    u = session.get('compras_user', '')
    if not u:
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db(); c = conn.cursor()

    # 1) Movimientos Entrada duplicados · mismo material+lote+cantidad,
    # < 10 min entre sí, en los últimos 4 días.
    # FIX 30-may-2026 · audit Planta · ANTES usaba julianday() (SQLite-only ·
    # rompía en PostgreSQL). Ahora: cutoff 4d en Python (string ISO) y la
    # ventana de 10 min entre el par se filtra en Python.
    from datetime import timedelta as _td7, datetime as _dt7
    _cut4d = (_dt7.now() - _td7(days=4)).isoformat()
    _pairs = c.execute(
        """SELECT m1.id, m2.id, m1.material_id, m1.material_nombre,
                  m1.lote, m1.cantidad, m1.fecha, m2.fecha
           FROM movimientos m1
           JOIN movimientos m2
             ON m1.material_id = m2.material_id
            AND m1.lote = m2.lote
            AND m1.cantidad = m2.cantidad
            AND m1.tipo = m2.tipo
            AND m1.id < m2.id
           WHERE m1.tipo = 'Entrada'
             AND m1.fecha >= ? AND m2.fecha >= ?
           ORDER BY m1.material_id, m1.fecha""",
        (_cut4d, _cut4d),
    ).fetchall()
    duplicados = []
    for r in _pairs:
        try:
            _f1 = _dt7.fromisoformat(str(r[6]))
            _f2 = _dt7.fromisoformat(str(r[7]))
            if abs((_f2 - _f1).total_seconds()) >= 600:  # >10 min · no es reintento
                continue
        except (ValueError, TypeError):
            continue  # fecha mal formada · omitir del reporte
        duplicados.append({
            'mov_id_1': r[0], 'mov_id_2': r[1], 'material_id': r[2],
            'material_nombre': r[3], 'lote': r[4], 'cantidad': r[5],
            'fecha_1': r[6], 'fecha_2': r[7],
        })

    # 2) Producciones iniciadas sin terminar
    colg_rows = c.execute(
        """SELECT id, producto, fecha_programada, inicio_real_at, estado,
                  COALESCE(inventario_descontado_at,'')
           FROM produccion_programada
           WHERE inicio_real_at IS NOT NULL
             AND fin_real_at IS NULL
             AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
           ORDER BY inicio_real_at DESC""",
    ).fetchall()
    colgadas = [{
        'id': r[0], 'producto': r[1], 'fecha_programada': r[2],
        'inicio_real_at': r[3], 'estado': r[4],
        'inventario_descontado': bool(r[5]),
    } for r in colg_rows]

    # 3) Lotes con stock negativo (tolerancia float -0.01)
    neg_rows = c.execute(
        """SELECT material_id, lote,
                  SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stock
           FROM movimientos
           GROUP BY material_id, lote
           HAVING stock < -0.01
           ORDER BY stock ASC""",
    ).fetchall()
    negativos = [{
        'material_id': r[0], 'lote': r[1], 'stock_neto': round(r[2], 2),
    } for r in neg_rows]

    todo_ok = not duplicados and not negativos
    return jsonify({
        'fecha_revision': datetime.now().isoformat(),
        'resumen': {
            'movimientos_entrada_duplicados': len(duplicados),
            'producciones_colgadas': len(colgadas),
            'lotes_stock_negativo': len(negativos),
        },
        'planta_ok': todo_ok,
        'duplicados_entrada': duplicados,
        'producciones_colgadas': colgadas,
        'lotes_stock_negativo': negativos,
        'nota': ('Si planta_ok=true, el incidente no dejó inconsistencias. '
                 'duplicados_entrada → anular el mov más nuevo con '
                 '/api/movimientos/<id>/anular. producciones_colgadas suele '
                 'ser estado intermedio legítimo (revisar manualmente).'),
    })


@bp.route('/api/lotes/cuarentena', methods=['GET'])
def lotes_cuarentena():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT m.id, m.material_id, m.material_nombre, m.lote, m.cantidad,
                      m.fecha, m.proveedor, m.numero_factura, m.numero_oc, m.observaciones,
                      mp.nombre_inci, m.estado_lote
               FROM movimientos m
               LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
               WHERE m.estado_lote IN ('CUARENTENA','CUARENTENA_EXTENDIDA') AND m.tipo='Entrada'
               ORDER BY m.fecha DESC""")
    rows = c.fetchall()
    cols = ['id','codigo_mp','nombre','lote','cantidad','fecha','proveedor','numero_factura','numero_oc','observaciones','nombre_inci','estado_lote']
    return jsonify([dict(zip(cols,r)) for r in rows])

@bp.route('/api/lotes/retenido', methods=['GET'])
def lotes_retenido():
    """Lotes NO disponibles con stock físico: RECHAZADO / VENCIDO / BLOQUEADO.

    Complementa /api/lotes/cuarentena (que solo lista CUARENTENA/_EXTENDIDA).
    A1 (Sebastián 12-jun) excluyó estos estados de /api/lotes para no enmascarar
    quiebres (M5), PERO el material físico sigue en bodega y debe permanecer
    TRAZABLE: INVIMA Res. 2214/2021 exige documentar el rechazado/vencido hasta
    su disposición (devolución/destrucción) y el conteo físico debe CUADRAR
    contra lo que el sistema muestra. Read-only · no muta nada.

    Stock NETO por (material_id, lote) con el CASE canónico (cuenta Ajuste como
    entrada · M-bodega), filtro de estado UPPER-insensible (M23), umbral >0.01 (M21).
    """
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT m.material_id, MAX(m.material_nombre) as nombre, m.lote,
                        SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as stock_neto,
                        MAX(m.fecha_vencimiento) as fecha_vencimiento,
                        MAX(m.proveedor) as proveedor, MAX(m.numero_oc) as numero_oc,
                        MAX(m.numero_factura) as numero_factura,
                        UPPER(COALESCE(MAX(m.estado_lote),'')) as estado_lote,
                        COALESCE(MAX(mp.nombre_inci),'') as nombre_inci
                 FROM movimientos m LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
                 WHERE UPPER(COALESCE(m.estado_lote,'')) IN ('RECHAZADO','VENCIDO','BLOQUEADO')
                 GROUP BY m.material_id, m.lote
                 HAVING stock_neto > 0.01
                 ORDER BY estado_lote, nombre""")
    rows = c.fetchall()
    cols = ['codigo_mp','nombre','lote','cantidad','fecha_vencimiento','proveedor','numero_oc','numero_factura','estado_lote','nombre_inci']
    return jsonify([dict(zip(cols,r)) for r in rows])

@bp.route('/api/lotes/liberar', methods=['POST'])
def liberar_lote():
    # Equipo de Calidad (CALIDAD_USERS) y admins pueden liberar lotes —
    # antes solo era admins, lo que bloqueaba el flujo legítimo de QC.
    u, err, code = _require_qc()
    if err:
        return err, code
    d = request.json or {}
    mov_id = d.get('id')
    accion = (d.get('accion') or 'APROBAR').upper()
    if accion not in ('APROBAR','RECHAZAR'):
        return jsonify({'error': 'Accion debe ser APROBAR o RECHAZAR'}), 400
    nuevo_estado = 'VIGENTE' if accion == 'APROBAR' else 'RECHAZADO'
    conn = get_db(); c = conn.cursor()
    # Firma electrónica Part 11 §11.200 · liberar/rechazar un lote en cuarentena
    # es una decisión regulada INVIMA · exige re-autenticación (password+MFA)
    # vía /api/sign. APROBAR→meaning 'libera', RECHAZAR→meaning 'rechaza'.
    meaning = 'libera' if accion == 'APROBAR' else 'rechaza'
    sig_id = d.get('signature_id')
    if not _validar_e_sign(c, sig_id, record_table='movimientos', record_id=mov_id,
                           meaning=meaning, signer_username=u):
        return jsonify({
            'error': 'Firma electrónica requerida',
            'requiere_firma': True,
            'sign_meaning': meaning,
            'record_table': 'movimientos',
            'record_id': str(mov_id),
            'detail': f"Firmá vía POST /api/sign con meaning='{meaning}', "
                      f"record_table='movimientos', record_id='{mov_id}' y reenviá signature_id.",
        }), 400
    # INVIMA-FIX · 21-may-2026 · aceptar CUARENTENA_EXTENDIDA también
    # Antes: lotes en cuarentena extendida quedaban zombies (rowcount=0)
    c.execute(
        "UPDATE movimientos SET estado_lote=? "
        "WHERE id=? AND estado_lote IN ('CUARENTENA','CUARENTENA_EXTENDIDA')",
        (nuevo_estado, mov_id),
    )
    if c.rowcount == 0:
        return jsonify({'error': 'Lote no encontrado o ya procesado'}), 404
    c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                 VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
              (u, f'LOTE_{accion}', 'movimientos',
               str(mov_id), f'Lote {accion} · firma e-sign #{sig_id}', request.remote_addr))
    conn.commit()
    return jsonify({'message': f'Lote {accion.lower()}ado correctamente', 'estado': nuevo_estado,
                     'signature_id': sig_id})

# /api/trazabilidad/<lote> eliminado — duplicado de /api/trazabilidad/lote/<path:lote>
# que es más completo (también busca despachos). Mantener el de path por
# consistencia con lotes que contienen '/' o caracteres especiales.

# ── CONTEO CICLICO BDG-PRO-002 ──────────────────────────────────
# Soporta filtro por tipo_material (MP / Envase Primario / Envase Secundario /
# Empaque) — clave para inventario cíclico de E&E (envase y empaque).
# Sin filtro = todo. Filtro = solo materiales del tipo indicado.
@bp.route('/api/conteo/estanterias', methods=['GET'])
def conteo_estanterias():
    tipo = (request.args.get('tipo_material') or '').strip()
    conn = get_db(); c = conn.cursor()
    if tipo in ('MP', 'Envase Primario', 'Envase Secundario', 'Empaque'):
        c.execute("""SELECT COALESCE(NULLIF(m.estanteria,''),'Sin estanteria') as est,
                            COUNT(DISTINCT m.material_id) as total_mps,
                            SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as stock_total
                     FROM movimientos m
                     LEFT JOIN maestro_mps mp ON m.material_id = mp.codigo_mp
                     WHERE COALESCE(mp.tipo_material,'MP') = ?
                     GROUP BY est ORDER BY est""", (tipo,))
    else:
        c.execute("""SELECT COALESCE(NULLIF(estanteria,''),'Sin estanteria') as est,
                            COUNT(DISTINCT material_id) as total_mps,
                            SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_total
                     FROM movimientos GROUP BY est ORDER BY est""")
    rows = c.fetchall()
    return jsonify([
        {'estanteria': r[0], 'total_mps': r[1], 'stock_total': round(r[2] or 0, 1)}
        for r in rows
    ])

@bp.route('/api/conteo/materiales', methods=['GET'])
def conteo_materiales_estanteria():
    """Items a contar en una estanteria. Devuelve UN row por (MP, lote).

    Cambio 2026-04-27 (CEO): el conteo ciclico antes agrupaba por MP
    perdiendo precision. Ahora cada lote es su propia fila — el operario
    puede contar lote por lote (que es como estan fisicamente acomodados
    en bodega). Lotes con stock_sistema <= 0 se filtran (HAVING) — no
    aparecen en el conteo, lo cual responde a "que pasa con lotes a 0":
    naturalmente desaparecen del conteo y de Bodega MP filtrada por
    stock > 0. Si el usuario quiere borrarlos formalmente del kardex,
    el boton 'Eliminar' con motivo (en Stock por Lote y aqui en conteo)
    hace el hard-delete con audit_log.

    Cada row trae proveedor + posicion + fecha_venc para que el operario
    contextualice: "este lote es de Inchemical, posicion E1-01, vence
    2027-08-19".
    """
    est = request.args.get('estanteria', '')
    tipo = (request.args.get('tipo_material') or '').strip()
    conn = get_db(); c = conn.cursor()
    type_filter = (
        " AND COALESCE(mp.tipo_material,'MP') = ?"
        if tipo in ('MP', 'Envase Primario', 'Envase Secundario', 'Empaque')
        else ""
    )
    type_args = (tipo,) if type_filter else ()
    # FIX 24-may-2026 noche · agente auditó: conteo cíclico contaba TODOS
    # los movimientos sin excluir CUARENTENA/VENCIDO/RECHAZADO/AGOTADO.
    # Resultado: el operario contaba físicamente solo el stock VIGENTE
    # en la estantería pero el sistema le mostraba stock_sistema con
    # lotes bloqueados incluidos → diferencia artificial siempre.
    # Ahora exclusión alineada con _get_mp_stock canónico.
    estado_lote_filter = (
        " AND (m.estado_lote IS NULL OR UPPER(COALESCE(m.estado_lote,'')) "
        "      NOT IN ('CUARENTENA','CUARENTENA_EXTENDIDA','RECHAZADO','VENCIDO','AGOTADO','BLOQUEADO'))"
    )
    if est and est != 'Sin estanteria':
        c.execute(f"""SELECT m.material_id, MAX(m.material_nombre) as material_nombre,
                            COALESCE(mp.nombre_inci,'') as inci,
                            COALESCE(mp.precio_referencia,0) as precio_ref,
                            SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as stock_sistema,
                            MAX(m.estanteria) as estanteria,
                            COALESCE(mp.tipo_material,'MP') as tipo_material,
                            COALESCE(m.lote,'') as lote,
                            COALESCE(MAX(m.proveedor),'') as proveedor,
                            COALESCE(MAX(m.posicion),'') as posicion,
                            COALESCE(MAX(m.fecha_vencimiento),'') as fecha_vencimiento
                     FROM movimientos m
                     LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
                     WHERE m.estanteria=?{type_filter}{estado_lote_filter}
                     GROUP BY m.material_id, m.lote, mp.codigo_mp HAVING stock_sistema > 0.01
                     ORDER BY material_nombre, m.lote""", (est,) + type_args)
    else:
        c.execute(f"""SELECT m.material_id, MAX(m.material_nombre) as material_nombre,
                            COALESCE(mp.nombre_inci,'') as inci,
                            COALESCE(mp.precio_referencia,0) as precio_ref,
                            SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as stock_sistema,
                            '' as estanteria,
                            COALESCE(mp.tipo_material,'MP') as tipo_material,
                            COALESCE(m.lote,'') as lote,
                            COALESCE(MAX(m.proveedor),'') as proveedor,
                            COALESCE(MAX(m.posicion),'') as posicion,
                            COALESCE(MAX(m.fecha_vencimiento),'') as fecha_vencimiento
                     FROM movimientos m
                     LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp
                     WHERE (m.estanteria IS NULL OR m.estanteria=''){type_filter}{estado_lote_filter}
                     GROUP BY m.material_id, m.lote, mp.codigo_mp HAVING stock_sistema > 0.01
                     ORDER BY material_nombre, m.lote""", type_args)
    rows = c.fetchall()
    cols = ['codigo_mp','nombre','inci','precio_ref','stock_sistema',
            'estanteria','tipo_material','lote','proveedor','posicion',
            'fecha_vencimiento']
    return jsonify([dict(zip(cols, r)) for r in rows])

@bp.route('/api/conteo/iniciar', methods=['POST'])
def conteo_iniciar():
    u, err, code = _require_planta_write()
    if err:
        return err, code
    d = request.json or {}
    est = d.get('estanteria', '')
    responsable = d.get('responsable', session.get('compras_user',''))
    from datetime import date
    conn = get_db(); c = conn.cursor()
    # ── Verificar si ya existe un conteo ABIERTO para esta estantería ──────────
    c.execute("SELECT id, numero FROM conteos_fisicos WHERE estanteria=? AND estado='Abierto' ORDER BY id DESC LIMIT 1", (est,))
    existing = c.fetchone()
    if existing:
        return jsonify({'conteo_id': existing[0], 'numero': existing[1],
                        'message': 'Conteo retomado', 'resuming': True})
    numero = 'CNT-' + date.today().strftime('%Y%m%d') + '-' + est.replace(' ','')[:6].upper()
    # Si el número ya existe (mismo día), agregar sufijo incremental
    c.execute("SELECT COUNT(*) FROM conteos_fisicos WHERE numero LIKE ?", (numero + '%',))
    suffix = c.fetchone()[0]
    if suffix > 0:
        numero = numero + f'-{suffix+1}'
    try:
        c.execute("INSERT INTO conteos_fisicos (numero,fecha_inicio,estado,responsable,estanteria,tipo_conteo) VALUES (?,datetime('now', '-5 hours'),'Abierto',?,?,'Ciclico')",
                  (numero, responsable, est))
        conteo_id = c.lastrowid
        conn.commit()
        return jsonify({'conteo_id': conteo_id, 'numero': numero, 'message': 'Conteo iniciado', 'resuming': False})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@bp.route('/api/conteo/programacion', methods=['GET'])
def conteo_programacion():
    """Retorna la programación cíclica automática.

    Modo MP (default, sin filtro): rotación por estantería.
    Modo E&E (?tipo_material=Empaque|Envase Primario|Envase Secundario):
        rotación de 3 ITEMS específicos por semana ISO. Como E&E no tiene
        ubicación física, en lugar de elegir estantería se eligen 3 ítems
        determinísticos para la semana — el equipo busca los 3 y los cuenta.

    Determinístico: la misma semana siempre devuelve los mismos 3 ítems.
    Rota por TODOS los ítems del tipo a lo largo del año.
    """
    tipo = (request.args.get('tipo_material') or '').strip()
    if tipo in ('Envase Primario', 'Envase Secundario', 'Empaque'):
        return _conteo_programacion_items(tipo)
    # Modo MP por estantería (legacy)
    from datetime import date, timedelta
    import math
    conn = get_db(); c = conn.cursor()
    # Obtener todas las estanterias con stock positivo desde movimientos
    c.execute("""SELECT COALESCE(NULLIF(estanteria,''),'Sin estanteria') as est
                 FROM movimientos
                 GROUP BY est
                 HAVING SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) > 0
                 ORDER BY est""")
    estanterias = [r[0] for r in c.fetchall()]
    if not estanterias:
        return jsonify({'semanas': [], 'total_estanterias': 0})
    n = len(estanterias)
    hoy = date.today()
    # Lunes de la semana actual
    lunes_actual = hoy - timedelta(days=hoy.weekday())
    semanas = []
    for delta in range(-1, 5):  # semana pasada + actual + próximas 4
        lunes = lunes_actual + timedelta(weeks=delta)
        semana_iso = lunes.isocalendar()[1]
        anio = lunes.isocalendar()[0]
        idx = (semana_iso - 1) % n
        est = estanterias[idx]
        # Verificar si ya hay conteo para esa semana + estantería
        fecha_fin = lunes + timedelta(days=6)
        c.execute("""SELECT id, numero, estado FROM conteos_fisicos
                     WHERE estanteria=? AND fecha_inicio BETWEEN ? AND ?
                     ORDER BY id DESC LIMIT 1""",
                  (est, lunes.isoformat(), fecha_fin.isoformat() + ' 23:59:59'))
        conteo = c.fetchone()
        semanas.append({
            'semana': semana_iso,
            'anio': anio,
            'lunes': lunes.isoformat(),
            'estanteria': est,
            'es_actual': delta == 0,
            'conteo_id': conteo[0] if conteo else None,
            'conteo_numero': conteo[1] if conteo else None,
            'conteo_estado': conteo[2] if conteo else 'Pendiente',
        })
    return jsonify({'semanas': semanas, 'total_estanterias': n, 'estanterias': estanterias})


def _conteo_programacion_items(tipo_material):
    """Programación cíclica para E&E: 3 ítems por semana, deterministas.

    El stock de Envase y Empaque vive en `maestro_mee` (NO en maestro_mps).
    No tiene localización física específica como las MPs, por eso la rotación
    es por ÍTEM en lugar de por estantería.

    Algoritmo:
      1. Lista ordenada de items activos en maestro_mee filtrados por
         categoría según el tipo solicitado:
           'Envase Primario'   → categoria LIKE '%envase%'  (frasco, tubo, gotero)
           'Envase Secundario' → categoria LIKE '%envase%'  (caja, display)
           'Empaque'           → categoria LIKE '%empaque%' (etiqueta, sello)
         Si solo hay un valor 'Envase' genérico, los 3 botones muestran lo
         mismo (es lo que el user pidió: rotar todos los E&E).
      2. Para una semana ISO N: índices [(3*N) % L, (3*N+1) % L, (3*N+2) % L]
         donde L = total ítems. Garantiza:
           - Mismos 3 ítems para la misma semana (determinístico).
           - Rotación completa: tras ⌈L/3⌉ semanas se han contado todos.
           - Sin solapamiento: 3 ítems distintos cada semana.
      3. Devuelve esquema compatible con UI: 'estanteria' lleva etiqueta
         sintética "E&E-<tipo>-S<sem>", 'items_programados' los códigos+nombres.
    """
    from datetime import date, timedelta

    conn = get_db(); c = conn.cursor()
    # Universo: items activos en maestro_mee. Filtramos por categoría según
    # el tipo solicitado. Si no hay distinción real en la categoría, devuelve
    # todos (es lo que el user quiere: que el jefe de planta vea cualquier
    # item de E&E rotando, no le importa la sub-clasificación rígida).
    if 'empaque' in tipo_material.lower():
        cat_pattern = '%empaque%'
    else:
        # Envase Primario o Envase Secundario → ambos toman de categoria
        # tipo "Envase" (sin distinguir, porque el catálogo no lo hace).
        cat_pattern = '%envase%'

    c.execute("""SELECT codigo, descripcion
                 FROM maestro_mee
                 WHERE LOWER(COALESCE(estado,''))='activo'
                   AND LOWER(COALESCE(categoria,'')) LIKE ?
                 ORDER BY codigo""", (cat_pattern,))
    items = c.fetchall()
    # Fallback: si la categoría no calza (ej: usuario pone categoría diferente
    # o solo tiene 'Otro'), devolver TODOS los items activos. El user prefiere
    # tener algo que rotar antes que un mensaje vacío.
    if not items:
        c.execute("""SELECT codigo, descripcion FROM maestro_mee
                     WHERE LOWER(COALESCE(estado,''))='activo'
                     ORDER BY codigo""")
        items = c.fetchall()

    if not items:
        # Diagnóstico contra maestro_mee (que es la tabla del stock de E&E)
        try:
            stats = c.execute("""
                SELECT COALESCE(NULLIF(TRIM(categoria),''),'(sin categoria)') AS cat,
                       COUNT(*) AS total
                FROM maestro_mee
                GROUP BY cat ORDER BY total DESC
            """).fetchall()
            cats_existentes = [{'tipo': r[0], 'total': r[1]} for r in stats]
            total_catalogo = sum(r['total'] for r in cats_existentes)
        except Exception:
            cats_existentes, total_catalogo = [], 0
        return jsonify({
            'semanas': [],
            'total_items': 0,
            'tipo_material': tipo_material,
            'mensaje': f"No hay items de E&E activos en maestro_mee para '{tipo_material}'.",
            'diagnostico': {
                'total_catalogo': total_catalogo,
                'sin_clasificar': 0,
                'tipos_existentes': cats_existentes,
                'accion_sugerida': (
                    "El catálogo de Envase y Empaque (maestro_mee) está vacío "
                    "o todos los items están en estado Inactivo. Ve a Planta → "
                    "Maestro de Envase y Empaque y agrega/activa los items que "
                    "tu equipo debe contar cíclicamente."
                ),
            }
        })

    L = len(items)
    hoy = date.today()
    lunes_actual = hoy - timedelta(days=hoy.weekday())
    semanas = []

    for delta in range(-1, 5):  # semana pasada + actual + próximas 4
        lunes = lunes_actual + timedelta(weeks=delta)
        semana_iso = lunes.isocalendar()[1]
        anio = lunes.isocalendar()[0]
        # Combinamos año + semana para que la rotación no se repita igual cada año
        seed = anio * 100 + semana_iso
        # 3 índices distintos, deterministas
        indices = [(3 * seed + offset) % L for offset in range(3)]
        # Si L < 3, evitar duplicados
        if L < 3:
            indices = list(range(L))
        items_semana = [
            {'codigo_mp': items[i][0], 'nombre': items[i][1]}
            for i in indices
        ]

        # Buscar conteo asociado de la semana (estanteria=etiqueta sintética)
        etiqueta = f"E&E-{tipo_material}-S{semana_iso:02d}"
        fecha_fin = lunes + timedelta(days=6)
        c.execute("""SELECT id, numero, estado FROM conteos_fisicos
                     WHERE estanteria=? AND fecha_inicio BETWEEN ? AND ?
                     ORDER BY id DESC LIMIT 1""",
                  (etiqueta, lunes.isoformat(), fecha_fin.isoformat() + ' 23:59:59'))
        conteo = c.fetchone()

        semanas.append({
            'semana': semana_iso,
            'anio': anio,
            'lunes': lunes.isoformat(),
            'estanteria': etiqueta,  # campo legacy compatible con UI
            'tipo_material': tipo_material,
            'items_programados': items_semana,
            'es_actual': delta == 0,
            'conteo_id': conteo[0] if conteo else None,
            'conteo_numero': conteo[1] if conteo else None,
            'conteo_estado': conteo[2] if conteo else 'Pendiente',
        })

    return jsonify({
        'semanas': semanas,
        'total_items': L,
        'tipo_material': tipo_material,
        'modo': 'items',  # vs 'estanteria' del modo MP
    })


@bp.route('/api/conteo/<int:conteo_id>/guardar', methods=['POST'])
def conteo_guardar(conteo_id):
    # P0 audit 26-may-2026 · sin gate · operario podía guardar conteos sin sesión.
    _u, _err, _code = _require_planta_write()
    if _err:
        return _err, _code
    d = request.json or {}
    items = d.get('items', [])
    UMBRAL_ESCALA = 0.05  # 5% -> escala a gerencia (BDG-PRO-002 num 8)
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT estado FROM conteos_fisicos WHERE id=?", (conteo_id,))
    row = c.fetchone()
    if not row or row[0] != 'Abierto':
        return jsonify({'error': 'Conteo no encontrado o ya cerrado'}), 400

    # FIX 24-may-2026 noche · validar NaN/Infinity/negativo en stock_fisico.
    # El trigger BD bloquea negativo pero NO NaN · UI puede mandar valores
    # malos por error del operario.
    import math as _math
    items_con_diff = 0
    items_invalidos = []
    for item in items:
        codigo = item.get('codigo_mp','')
        _lote_it = (item.get('lote','') or '')
        # Cero-error (Sebastian 12-jun): si el ajuste de este item YA se aplico,
        # NO re-escribir su fila. El INSERT OR REPLACE de abajo resetearia
        # ajuste_aplicado/aprobado_gerencia (no estan en la lista de columnas) ->
        # el item volveria a 'no ajustado' y cerrar/ajustar lo aplicaria una 2da
        # vez (doble entrada/salida en el kardex). Una vez aplicado, queda fijo.
        try:
            _ya = c.execute(
                "SELECT COALESCE(ajuste_aplicado,0) FROM conteo_items "
                "WHERE conteo_id=? AND codigo_mp=? AND COALESCE(lote,'')=?",
                (conteo_id, codigo, _lote_it)).fetchone()
            if _ya and _ya[0]:
                continue
        except Exception:
            pass
        try:
            stock_sis = float(item.get('stock_sistema', 0))
        except (TypeError, ValueError):
            stock_sis = 0
        stock_fis = item.get('stock_fisico')
        if stock_fis is None or stock_fis == '': continue
        try:
            stock_fis = float(stock_fis)
        except (TypeError, ValueError):
            items_invalidos.append({'codigo': codigo, 'razon': 'stock_fisico no numérico'})
            continue
        if _math.isnan(stock_fis) or _math.isinf(stock_fis):
            items_invalidos.append({'codigo': codigo, 'razon': 'stock_fisico NaN/Infinity'})
            continue
        if stock_fis < 0:
            items_invalidos.append({'codigo': codigo, 'razon': 'stock_fisico negativo'})
            continue
        if stock_fis > 1_000_000_000:
            items_invalidos.append({'codigo': codigo, 'razon': 'stock_fisico absurdamente grande'})
            continue
        diff = stock_fis - stock_sis
        precio_ref = float(item.get('precio_ref', 0))
        valor_diff = abs(diff / 1000) * precio_ref  # diff en g, precio en /kg
        # FIX-B6 13-may-2026: antes pct_diff=0 cuando stock_sis=0 · cualquier
        # cantidad encontrada físicamente sin sistema previo evadía el
        # threshold gerencia. Ahora si stock_sis=0 Y diff != 0 → escalación
        # automática (deviación infinita = siempre supera 5%).
        if stock_sis > 0:
            pct_diff = abs(diff / stock_sis)
        else:
            pct_diff = 1.0 if abs(diff) > 0.01 else 0
        requiere_gerencia = 1 if pct_diff > UMBRAL_ESCALA else 0
        causa = item.get('causa_diferencia', '')
        if abs(diff) > 0: items_con_diff += 1
        c.execute("""INSERT OR REPLACE INTO conteo_items
                     (conteo_id,codigo_mp,nombre_mp,stock_sistema,stock_fisico,diferencia,
                      estanteria,causa_diferencia,valor_diferencia,requiere_gerencia,observaciones,lote)
                     VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (conteo_id, codigo, item.get('nombre',''), stock_sis, stock_fis, diff,
                   item.get('estanteria',''), causa, round(valor_diff,0), requiere_gerencia,
                   item.get('observaciones',''), item.get('lote','') or ''))
    c.execute("UPDATE conteos_fisicos SET items_diferencia=?,total_items=? WHERE id=?",
              (items_con_diff, len(items), conteo_id))
    try:
        c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (session.get('compras_user',''), 'GUARDAR_CONTEO', 'conteos_fisicos',
                   str(conteo_id),
                   f'Conteo #{conteo_id}: {len(items)} items, {items_con_diff} con diferencia',
                   request.remote_addr if request else ''))
    except sqlite3.OperationalError:
        pass
    conn.commit()
    # Sebastian 7-may-2026: devolver items con sus IDs para que la UI pueda
    # llamar /ajustar(item_id) directo en cada fila (botón "Aplicar ajuste")
    saved_items = []
    try:
        for r in c.execute("""
            SELECT id, codigo_mp, COALESCE(lote,'') as lote, diferencia,
                   COALESCE(requiere_gerencia,0), COALESCE(aprobado_gerencia,0),
                   COALESCE(ajuste_aplicado,0)
            FROM conteo_items WHERE conteo_id=?
        """, (conteo_id,)).fetchall():
            saved_items.append({
                'id': r[0], 'codigo_mp': r[1], 'lote': r[2],
                'diferencia': float(r[3] or 0),
                'requiere_gerencia': bool(r[4]),
                'aprobado_gerencia': bool(r[5]),
                'ajuste_aplicado': bool(r[6]),
            })
    except Exception:
        pass
    return jsonify({
        'message': 'Conteo guardado',
        'items_con_diferencia': items_con_diff,
        'items_invalidos': items_invalidos,
        'items': saved_items,
    })

@bp.route('/api/conteo/<int:conteo_id>/cerrar', methods=['POST'])
def conteo_cerrar(conteo_id):
    """Cierra un conteo físico aplicando ajustes automáticos.

    Flujo deseado por el dueño del negocio:
      - Diferencias <5% del stock sistema → AUTO-AJUSTE: se inserta movimiento
        de Entrada/Salida que sincroniza el kardex con la realidad física.
      - Diferencias >=5% → NO se aplica ajuste. Se marca como pendiente de
        gerencia y se genera ALERTA en panel (BDG-PRO-002 num 8). Gerencia
        revisa el caso (puede ser hurto, mal pesaje, error de fórmula) y
        aprueba manualmente con /api/conteo/<id>/ajustar.

    Trazabilidad: cada auto-ajuste queda en `movimientos` con lote
    'AJUSTE-CICLICO-<conteo_id>' y observaciones que indican causa.
    """
    # FIX 1-jun-2026 (audit) · era el único write del flujo de conteo SIN gate de
    # auth → petición no autenticada podía mutar el kardex y el audit registraba
    # 'sistema'. Ahora exige planta-write (igual que conteo_iniciar/ajustar).
    u, err, code = _require_planta_write()
    if err:
        return err, code
    user = u
    conn = get_db(); c = conn.cursor()

    c.execute("SELECT estado FROM conteos_fisicos WHERE id=?", (conteo_id,))
    cf = c.fetchone()
    if not cf:
        return jsonify({'error': 'Conteo no encontrado'}), 404
    if cf[0] == 'Cerrado':
        return jsonify({'error': 'El conteo ya estaba cerrado'}), 400

    # ── Items con diferencia ────────────────────────────────────────────────
    c.execute("""SELECT id, codigo_mp, nombre_mp, stock_sistema, stock_fisico,
                        diferencia, estanteria, causa_diferencia, valor_diferencia,
                        requiere_gerencia, aprobado_gerencia, ajuste_aplicado
                 FROM conteo_items
                 WHERE conteo_id=? AND COALESCE(diferencia,0) <> 0""", (conteo_id,))
    items_con_diff = c.fetchall()

    auto_ajustados = []
    pendientes_gerencia_lista = []

    try:
        for it in items_con_diff:
            (it_id, codigo, nombre, stock_sis, stock_fis, diff, estant,
             causa, valor, req_ger, aprob_ger, ya_ajustado) = it
            if ya_ajustado:
                continue
            diff = float(diff or 0)
            if diff == 0:
                continue
            tipo_mov = 'Entrada' if diff > 0 else 'Salida'

            # Sebastian 12-jun: el jefe de produccion y los operarios ajustan el
            # inventario SIN aprobacion de Gerencia ni correo. Antes los items con
            # diferencia >5% NO se auto-ajustaban (quedaban pendientes + se enviaba
            # email a gerencia). Ahora se auto-ajustan TODOS; la trazabilidad queda
            # en audit_log (operario + causa + monto). El flag requiere_gerencia se
            # conserva solo como marca informativa para reportes.

            # Auto-ajuste para diferencias menores
            # Sebastian 7-may-2026: aplicar al lote REAL del conteo (no
            # al sintético) para que la query por lote en Bodega refleje
            # el cambio. Fallback a 'AJUSTE-CICLICO-<id>' si no hay lote.
            lote_real = ''
            try:
                lote_row = c.execute(
                    "SELECT COALESCE(lote,'') FROM conteo_items WHERE id=?",
                    (it_id,)
                ).fetchone()
                lote_real = (lote_row[0] if lote_row else '') or ''
            except Exception:
                lote_real = ''
            lote_obj = lote_real or f'AJUSTE-CICLICO-{conteo_id}'
            obs = (f"Ajuste automático conteo cíclico #{conteo_id} | "
                   f"Causa: {causa or 'no indicada'} | Cerrado por: {user}")
            # FIX 24-may noche · Sebastián: 'si modifican aquí se refleje
            # en bodega de MP y MEE'. Antes solo INSERT a movimientos (MP).
            # Si el código es MEE, el INSERT en MP era invisible y MEE
            # quedaba sin ajustar. Ahora detectar tipo material:
            #  - existe en maestro_mee → INSERT movimientos_mee + helper
            #    canónico que actualiza stock_actual atómicamente.
            #  - sino → INSERT movimientos (kardex MP).
            es_mee = False
            try:
                mee_row = c.execute(
                    "SELECT 1 FROM maestro_mee WHERE UPPER(TRIM(codigo)) = UPPER(TRIM(?))",
                    (codigo,),
                ).fetchone()
                es_mee = bool(mee_row)
            except Exception:
                es_mee = False
            if es_mee:
                # Aplicar al kardex MEE (movimientos_mee + stock_actual cache).
                try:
                    from inventario_helpers import aplicar_movimiento_mee
                    aplicar_movimiento_mee(
                        conn, codigo, tipo_mov, abs(diff),
                        observaciones=obs, responsable=user,
                        lote_ref=str(conteo_id), batch_ref=lote_obj or '',
                    )
                except Exception as _e_mee:
                    # Fallback: INSERT directo en movimientos_mee si helper falla
                    try:
                        c.execute(
                            """INSERT INTO movimientos_mee
                                 (mee_codigo, descripcion, tipo, cantidad,
                                  fecha, observaciones, lote_ref, batch_ref, responsable, anulado)
                               VALUES (?,?,?,?,datetime('now', '-5 hours'),?,?,?,?,0)""",
                            (codigo, nombre, tipo_mov, abs(diff),
                             obs, str(conteo_id), lote_obj or '', user),
                        )
                    except Exception:
                        pass
            else:
                # Kardex MP (movimientos) · comportamiento original
                c.execute("""INSERT INTO movimientos
                             (material_id, material_nombre, cantidad, tipo, fecha,
                              observaciones, lote, estanteria, estado_lote, operador)
                             VALUES (?,?,?,?,datetime('now', '-5 hours'),?,?,?,'VIGENTE',?)""",
                          (codigo, nombre, abs(diff), tipo_mov, obs,
                           lote_obj, estant or '', user))
            c.execute("UPDATE conteo_items SET ajuste_aplicado=1 WHERE id=?", (it_id,))
            try:
                c.execute("""INSERT INTO audit_log
                             (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                             VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                          (user, 'AJUSTE_INVENTARIO_AUTO', 'conteo_items', str(it_id),
                           f'{"MEE" if es_mee else "MP"}:{codigo} Diff:{diff}g Auto:<5% Causa:{causa or "n/a"}',
                           request.remote_addr if request else ''))
            except sqlite3.OperationalError:
                pass  # audit_log puede no existir en versiones viejas
            auto_ajustados.append({
                'item_id': it_id, 'codigo_mp': codigo, 'nombre': nombre,
                'diferencia_g': diff, 'tipo': tipo_mov,
                'kardex': 'MEE' if es_mee else 'MP',
            })

        c.execute("""UPDATE conteos_fisicos
                     SET estado='Cerrado', fecha_cierre=datetime('now', '-5 hours')
                     WHERE id=?""", (conteo_id,))
        conn.commit()
    except Exception as _e:
        conn.rollback()
        __import__('logging').getLogger('inventario').error(
            "conteo_cerrar(%s) FALLÓ — rollback aplicado: %s",
            conteo_id, _e, exc_info=True
        )
        return jsonify({
            'error': 'Falla transaccional al cerrar conteo',
            'detalle': str(_e),
            'rollback': 'aplicado — ningún ajuste se persistió',
        }), 500

    # Email de alerta a gerencia si hay pendientes (best-effort, no bloquea).
    # Se envía a EMAIL_GERENCIA (env var, separado del buzón de facturación).
    # Soporta múltiples destinatarios separados por coma.
    # Si EMAIL_GERENCIA no está configurado, fallback a EMAIL_REMITENTE para
    # no perder el alerta — pero el operador debe configurarlo en Render.
    if pendientes_gerencia_lista:
        try:
            from notificaciones import SistemaNotificaciones
            import os as _os
            sn = SistemaNotificaciones()
            if sn.email_remitente and sn.contraseña:
                gerencia_raw = _os.environ.get('EMAIL_GERENCIA', '').strip()
                if gerencia_raw:
                    destinatarios = [e.strip() for e in gerencia_raw.split(',')
                                     if e.strip() and '@' in e.strip()]
                else:
                    destinatarios = [sn.email_remitente]
                total_valor = sum(p.get('valor_diferencia') or 0 for p in pendientes_gerencia_lista)
                lista_html = ''.join([
                    f"<li><strong>{p['codigo_mp']}</strong> — {p['nombre']}: "
                    f"diff <strong>{p['diferencia_g']:+.0f}g</strong> "
                    f"(${(p.get('valor_diferencia') or 0):,.0f}) — {p.get('causa') or 'sin causa'}</li>"
                    for p in pendientes_gerencia_lista[:20]
                ])
                body = f"""<html><body style="font-family:Arial,sans-serif">
                <h2 style="color:#c62828">Alerta — Conteo cíclico requiere aprobación Gerencia</h2>
                <p>Conteo <strong>#{conteo_id}</strong> cerrado por <strong>{user}</strong>.</p>
                <p>{len(pendientes_gerencia_lista)} item(s) con diferencia &gt;5% no fueron ajustados
                automáticamente. Valor total estimado: <strong>${total_valor:,.0f}</strong>.</p>
                <ul>{lista_html}</ul>
                <p>Ingresa a /planta y aprueba/rechaza cada item para que el ajuste se aplique al kardex.</p>
                </body></html>"""
                sn.enviar_en_background(
                    sn._enviar_email,
                    asunto=f"[HHA] Conteo cíclico #{conteo_id}: {len(pendientes_gerencia_lista)} items requieren Gerencia",
                    body=body,
                    destinatarios=destinatarios,
                )
        except Exception as _e_mail:
            __import__('logging').getLogger('inventario').error(
                "Email alerta gerencia conteo %s falló: %s", conteo_id, _e_mail
            )

    msg_parts = [f'Conteo #{conteo_id} cerrado.']
    if auto_ajustados:
        msg_parts.append(f'{len(auto_ajustados)} ajuste(s) automático(s) aplicado(s) al kardex.')
    if pendientes_gerencia_lista:
        msg_parts.append(
            f'ATENCIÓN: {len(pendientes_gerencia_lista)} item(s) con diferencia >5% '
            f'pendientes de aprobación Gerencia General (BDG-PRO-002).')
    return jsonify({
        'message': ' '.join(msg_parts),
        'auto_ajustados': auto_ajustados,
        'pendientes_gerencia': pendientes_gerencia_lista,
        'total_items_ajustados': len(auto_ajustados),
        'total_pendientes': len(pendientes_gerencia_lista),
    })


@bp.route('/api/conteo/alertas-gerencia', methods=['GET'])
def conteo_alertas_gerencia():
    """INFORME de ajustes grandes (>5%) para revisión de Gerencia.

    Sebastian 12-jun: los operarios/jefe ajustan SIN aprobación previa. Este
    endpoint dejó de ser una cola de aprobación y pasó a ser un INFORME: muestra
    los ajustes grandes YA APLICADOS (con quién los hizo + causa + monto + cuándo)
    para que Gerencia los revise cuando quiera, sin bloquear ni aprobar.
    `pendientes` queda por compatibilidad (normalmente vacío).
    """
    user = session.get('compras_user','')
    if not user:
        return jsonify({'error': 'No autenticado'}), 401
    conn = get_db(); c = conn.cursor()
    # Pendientes (compat · normalmente vacío ya que se auto-aplican)
    c.execute("""SELECT cf.id AS conteo_id, cf.numero, cf.estanteria, cf.fecha_cierre,
                        ci.id AS item_id, ci.codigo_mp, ci.nombre_mp,
                        ci.stock_sistema, ci.stock_fisico, ci.diferencia,
                        ci.causa_diferencia, ci.valor_diferencia,
                        ci.aprobado_gerencia, ci.aprobado_gerencia_por,
                        ci.ajuste_aplicado
                 FROM conteo_items ci
                 JOIN conteos_fisicos cf ON ci.conteo_id = cf.id
                 WHERE ci.requiere_gerencia=1 AND ci.ajuste_aplicado=0
                 ORDER BY ABS(ci.valor_diferencia) DESC LIMIT 200""")
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    # INFORME · ajustes grandes YA aplicados (quién + causa + monto + cuándo)
    c.execute("""SELECT cf.id AS conteo_id, cf.numero, cf.estanteria, cf.fecha_cierre,
                        ci.id AS item_id, ci.codigo_mp, ci.nombre_mp,
                        ci.stock_sistema, ci.stock_fisico, ci.diferencia,
                        ci.causa_diferencia, ci.valor_diferencia,
                        ci.aprobado_gerencia_por AS aplicado_por
                 FROM conteo_items ci
                 JOIN conteos_fisicos cf ON ci.conteo_id = cf.id
                 WHERE ci.requiere_gerencia=1 AND ci.ajuste_aplicado=1
                 ORDER BY cf.fecha_cierre DESC, ABS(ci.valor_diferencia) DESC LIMIT 200""")
    cols2 = [d[0] for d in c.description]
    aplicados = [dict(zip(cols2, r)) for r in c.fetchall()]
    total_valor = sum(r.get('valor_diferencia') or 0 for r in rows)
    return jsonify({
        'pendientes': rows,
        'total': len(rows),
        'total_valor_diferencia': round(total_valor, 0),
        'aplicados': aplicados,
        'total_aplicados': len(aplicados),
        'total_valor_aplicados': round(sum(r.get('valor_diferencia') or 0 for r in aplicados), 0),
    })

@bp.route('/api/conteo/<int:conteo_id>/ajustar', methods=['POST'])
def conteo_ajustar(conteo_id):
    # P0 audit 26-may-2026 · sin gate sesión, el check ADMIN_USERS más abajo
    # se bypasea con user='' (no está en ADMIN_USERS, pero igual cae al
    # else que mete movimientos sin auditoría real).
    _u_sess, _err_s, _code_s = _require_session()
    if _err_s:
        return _err_s, _code_s
    user = session.get('compras_user','')
    if not user:
        return jsonify({'error': 'sesión inválida'}), 401
    d = request.json or {}
    item_id = d.get('item_id')
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT ci.*, cf.estado FROM conteo_items ci JOIN conteos_fisicos cf ON ci.conteo_id=cf.id WHERE ci.id=?", (item_id,))
    item = c.fetchone()
    if not item:
        return jsonify({'error': 'Item no encontrado'}), 404
    cols = [desc[0] for desc in c.description]
    it = dict(zip(cols, item))
    # FIX-B2 13-may-2026: idempotencia. Antes el endpoint NO chequeaba
    # ajuste_aplicado al entrar, así que un doble-click del operario (o un
    # reintento tras timeout) duplicaba el INSERT en movimientos · stock
    # quedaba con −2× del delta real. Ahora si ya se aplicó devolvemos 409
    # como no-op idempotente. Atomic claim vía UPDATE-WHERE para evitar
    # race condition (2 requests concurrentes).
    if it.get('ajuste_aplicado'):
        return jsonify({
            'error': 'Ajuste ya aplicado para este item · operación idempotente',
            'item_id': item_id,
            'aplicado_anterior': True,
        }), 409
    # Sebastian 12-jun: el jefe de produccion y los operarios ajustan SIN
    # aprobacion de Gerencia (antes >5% exigia ADMIN -> 403). Ahora cualquier
    # usuario con acceso al conteo aplica el ajuste; se registra quien lo hizo
    # (aprobado_gerencia_por=user) + audit_log -> trazabilidad sin bloqueo.
    # Los ajustes grandes quedan en el INFORME /api/conteo/alertas-gerencia.
    if it['requiere_gerencia'] and not it['aprobado_gerencia']:
        c.execute("UPDATE conteo_items SET aprobado_gerencia=1,aprobado_gerencia_por=? WHERE id=?", (user, item_id))
    diff = float(it['diferencia'])
    if diff == 0:
        return jsonify({'message': 'Sin diferencia, no se requiere ajuste'})

    # Atomic claim: marcar ajuste_aplicado=1 en mismo statement con
    # WHERE ajuste_aplicado=0. Si rowcount=0, otro request ya lo procesó.
    c.execute("UPDATE conteo_items SET ajuste_aplicado=1 WHERE id=? AND COALESCE(ajuste_aplicado,0)=0",
              (item_id,))
    if c.rowcount == 0:
        conn.rollback()
        return jsonify({
            'error': 'Otro request procesó este ajuste en paralelo · operación idempotente',
            'item_id': item_id,
        }), 409
    tipo_mov = 'Entrada' if diff > 0 else 'Salida'
    # Sebastian 7-may-2026: usar el LOTE REAL del conteo (it['lote']) para
    # que el movimiento afecte el kardex del lote correcto y Bodega lo
    # refleje. Antes usábamos 'AJUSTE-<id>' (lote sintético) y la query por
    # lote en bodega seguía mostrando el stock viejo del lote original.
    # Si el conteo no tiene lote (material agregado sin lote), fallback al
    # sintético 'AJUSTE-<id>' para no perder trazabilidad.
    lote_objetivo = (it.get('lote') or '').strip() or f'AJUSTE-{conteo_id}'
    obs = (f'Ajuste inventario ciclico #{conteo_id} - '
           f'{it.get("causa_diferencia","Sin causa")} - Aprobado: {user}')
    # FIX 24-may noche · si el item es MEE, ajustar en movimientos_mee.
    es_mee_a = False
    try:
        mee_row_a = c.execute(
            "SELECT 1 FROM maestro_mee WHERE UPPER(TRIM(codigo)) = UPPER(TRIM(?))",
            (it['codigo_mp'],),
        ).fetchone()
        es_mee_a = bool(mee_row_a)
    except Exception:
        es_mee_a = False
    if es_mee_a:
        try:
            from inventario_helpers import aplicar_movimiento_mee
            aplicar_movimiento_mee(
                conn, it['codigo_mp'], tipo_mov, abs(diff),
                observaciones=obs, responsable=user,
                lote_ref=str(conteo_id), batch_ref=lote_objetivo or '',
            )
        except Exception:
            try:
                c.execute(
                    """INSERT INTO movimientos_mee
                         (mee_codigo, descripcion, tipo, cantidad,
                          fecha, observaciones, lote_ref, batch_ref, responsable, anulado)
                       VALUES (?,?,?,?,datetime('now', '-5 hours'),?,?,?,?,0)""",
                    (it['codigo_mp'], it['nombre_mp'], tipo_mov, abs(diff),
                     obs, str(conteo_id), lote_objetivo or '', user),
                )
            except Exception:
                pass
    else:
        c.execute("""INSERT INTO movimientos (material_id,material_nombre,cantidad,tipo,fecha,observaciones,lote,estanteria,estado_lote,operador)
                     VALUES (?,?,?,?,datetime('now', '-5 hours'),?,?,?,'VIGENTE',?)""",
                  (it['codigo_mp'], it['nombre_mp'], abs(diff), tipo_mov, obs,
                   lote_objetivo, it.get('estanteria',''), user))
    # ajuste_aplicado=1 ya seteado en el atomic claim arriba (FIX-B2)
    c.execute("INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha) VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))",
              (user, 'AJUSTE_INVENTARIO', 'conteo_items', str(item_id),
               f'MP:{it["codigo_mp"]} Diff:{diff}g Lote:{lote_objetivo} Causa:{it.get("causa_diferencia","")}',
               request.remote_addr))
    conn.commit()
    # Stock actual post-ajuste (suma todos movimientos del lote)
    stock_post = 0.0
    try:
        row = c.execute("""
            SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad
                                     WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad
                                     ELSE 0 END), 0)
            FROM movimientos WHERE material_id=? AND lote=?
        """, (it['codigo_mp'], lote_objetivo)).fetchone()
        stock_post = float(row[0] or 0)
    except Exception:
        pass
    return jsonify({
        'message': f'Ajuste aplicado: {tipo_mov} de {abs(diff):.0f}g para {it["nombre_mp"]} (lote {lote_objetivo}). Stock post-ajuste del lote: {stock_post:.0f}g',
        'lote_ajustado': lote_objetivo,
        'tipo_movimiento': tipo_mov,
        'cantidad_g': abs(diff),
        'stock_lote_post': stock_post,
    })

@bp.route('/api/conteo/historial', methods=['GET'])
def conteo_historial():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT cf.id, cf.numero, cf.estanteria, cf.fecha_inicio, cf.fecha_cierre,
                        cf.estado, cf.responsable, cf.total_items, cf.items_diferencia,
                        COUNT(CASE WHEN ci.requiere_gerencia=1 THEN 1 END) as items_gerencia
                 FROM conteos_fisicos cf
                 LEFT JOIN conteo_items ci ON cf.id=ci.conteo_id
                 GROUP BY cf.id ORDER BY cf.fecha_inicio DESC LIMIT 50""")
    rows = c.fetchall()
    cols = ['id','numero','estanteria','fecha_inicio','fecha_cierre','estado','responsable','total_items','items_diferencia','items_gerencia']
    return jsonify([dict(zip(cols, r)) for r in rows])


@bp.route('/admin/conteo-rescate', methods=['GET'])
def admin_conteo_rescate():
    """RESCATE de conteo cíclico (Sebastián 6-jun-2026) · incidente "lo de ayer
    no sale". Muestra TODOS los conteos recientes con sus items, SIN filtro de
    fecha/semana/estado, para verificar que NADA se perdió y recuperar lo que los
    operarios ingresaron. Server-side (sin JS). Admin/Calidad."""
    u = session.get('compras_user', '')
    if not u:
        return redirect('/login?next=/admin/conteo-rescate')
    if u not in (ADMIN_USERS | CALIDAD_USERS):
        return Response('<div style="font-family:sans-serif;padding:40px;color:#991b1b">Solo Admin o Calidad pueden ver el rescate de conteo.</div>',
                        status=403, mimetype='text/html')
    import html as _hh
    dias = request.args.get('dias', '3')
    try:
        dias = max(1, min(60, int(dias)))
    except (TypeError, ValueError):
        dias = 3
    conn = get_db(); c = conn.cursor()
    # Fecha de corte calculada en Python (PG-safe · sin datetime() en el WHERE).
    corte = (datetime.utcnow() - timedelta(hours=5) - timedelta(days=dias)).strftime('%Y-%m-%d %H:%M:%S')
    # Conteos recientes (por fecha_inicio) SIN filtrar estado ni semana.
    c.execute(
        "SELECT id, COALESCE(numero,''), COALESCE(estanteria,''), COALESCE(fecha_inicio,''), "
        "COALESCE(fecha_cierre,''), COALESCE(estado,''), COALESCE(responsable,''), "
        "COALESCE(total_items,0), COALESCE(items_diferencia,0), COALESCE(tipo_conteo,'') "
        "FROM conteos_fisicos WHERE fecha_inicio >= ? "
        "ORDER BY fecha_inicio DESC", (corte,))
    conteos = c.fetchall()
    bloques = []
    tot_items = 0
    for ct in conteos:
        cid = ct[0]
        c.execute(
            "SELECT COALESCE(codigo_mp,''), COALESCE(nombre_mp,''), COALESCE(stock_sistema,0), "
            "COALESCE(stock_fisico,0), COALESCE(diferencia,0), COALESCE(lote,''), "
            "COALESCE(observaciones,''), COALESCE(ajuste_aplicado,0) "
            "FROM conteo_items WHERE conteo_id=? ORDER BY nombre_mp", (cid,))
        items = c.fetchall()
        tot_items += len(items)
        filas = ''.join(
            '<tr><td class="mono">' + _hh.escape(str(it[0])) + '</td>'
            '<td>' + _hh.escape(str(it[1])) + '</td>'
            '<td class="mono">' + _hh.escape(str(it[5])) + '</td>'
            '<td class="num">' + ('{:,.0f}'.format(it[2])) + '</td>'
            '<td class="num"><b>' + ('{:,.0f}'.format(it[3])) + '</b></td>'
            '<td class="num" style="color:' + ('#b91c1c' if it[4] else '#64748b') + '">' + ('{:,.0f}'.format(it[4])) + '</td>'
            '<td>' + _hh.escape(str(it[6])) + '</td>'
            '<td class="c">' + ('✓ aplicado' if it[7] else '—') + '</td></tr>'
            for it in items)
        if not filas:
            filas = '<tr><td colspan="8" class="muted">— sin items en este conteo —</td></tr>'
        estado_color = '#166534' if 'cerr' in (ct[5] or '').lower() else ('#854d0e' if 'abier' in (ct[5] or '').lower() else '#64748b')
        bloques.append(
            '<div class="conteo">'
            '<div class="ch"><div><b>Conteo ' + _hh.escape(str(ct[1] or ('#' + str(cid)))) + '</b> · Estantería ' + _hh.escape(str(ct[2] or '—')) +
            '</div><div>' + str(len(items)) + ' items · <span style="color:' + estado_color + ';font-weight:700">' + _hh.escape(str(ct[5] or '—')) + '</span></div></div>'
            '<div class="cm">Iniciado: ' + _hh.escape(str(ct[3])[:16]) + ' · Responsable: ' + _hh.escape(str(ct[6] or '—')) +
            (' · Cerrado: ' + _hh.escape(str(ct[4])[:16]) if ct[4] else '') + '</div>'
            '<table><thead><tr><th>Código</th><th>Material</th><th>Lote</th><th class="num">Sistema</th>'
            '<th class="num">Físico</th><th class="num">Dif.</th><th>Observación</th><th>Ajuste</th></tr></thead>'
            '<tbody>' + filas + '</tbody></table></div>')
    cuerpo = ''.join(bloques) if bloques else '<div class="muted" style="padding:30px;text-align:center">No hay conteos en los últimos ' + str(dias) + ' días.</div>'
    # Correcciones de Bodega MP (audit_log) — evidencia de que lo editado ayer
    # quedó REGISTRADO y APLICADO (los endpoints PUT actualizan todas las filas
    # del lote + commit). Acciones: ubicación, fecha venc, código lote, proveedor,
    # eliminar, y ajustes de inventario.
    corr_rows = []
    try:
        acciones = ('EDITAR_UBICACION_LOTE', 'EDITAR_FECHA_VENC_LOTE', 'EDITAR_CODIGO_LOTE',
                    'EDITAR_PROVEEDOR_LOTE', 'EDITAR_PROVEEDOR_MP', 'ELIMINAR_LOTE',
                    'AJUSTE_INVENTARIO', 'AJUSTE_INVENTARIO_CONTEO', 'AJUSTE_INVENTARIO_AUTO')
        ph = ','.join(['?'] * len(acciones))
        c.execute(
            "SELECT fecha, COALESCE(usuario,''), COALESCE(accion,''), COALESCE(detalle,'') "
            "FROM audit_log WHERE accion IN (" + ph + ") AND fecha >= ? "
            "ORDER BY fecha DESC LIMIT 400", tuple(acciones) + (corte,))
        corr_rows = c.fetchall()
    except Exception:
        corr_rows = []
    _accion_label = {
        'EDITAR_UBICACION_LOTE': '📍 Ubicación', 'EDITAR_FECHA_VENC_LOTE': '📅 Fecha venc.',
        'EDITAR_CODIGO_LOTE': '🏷 Código lote', 'EDITAR_PROVEEDOR_LOTE': '🏭 Proveedor lote',
        'EDITAR_PROVEEDOR_MP': '🏭 Proveedor MP', 'ELIMINAR_LOTE': '🗑 Eliminar lote',
        'AJUSTE_INVENTARIO': '⚖ Ajuste', 'AJUSTE_INVENTARIO_CONTEO': '⚖ Ajuste conteo',
        'AJUSTE_INVENTARIO_AUTO': '⚖ Ajuste auto'}
    corr_filas = ''.join(
        '<tr><td class="mono" style="white-space:nowrap">' + _hh.escape(str(cr[0])[:16]) + '</td>'
        '<td>' + _hh.escape(str(cr[1])) + '</td>'
        '<td>' + _hh.escape(_accion_label.get(cr[2], cr[2])) + '</td>'
        '<td style="font-size:11px;color:#475569">' + _hh.escape(str(cr[3])[:220]) + '</td></tr>'
        for cr in corr_rows)
    if not corr_filas:
        corr_filas = '<tr><td colspan="4" class="muted">— sin correcciones de Bodega MP en este período —</td></tr>'
    corr_html = (
        '<h2 style="color:#7c3aed;margin:22px 0 8px;font-size:18px">🛠 Correcciones en Bodega MP (últimos ' + str(dias) + ' días)</h2>'
        '<div class="sub" style="margin-bottom:10px">Cada corrección de lote (ubicación, fecha, código, proveedor) y cada ajuste de inventario queda auditado aquí. '
        'Esto confirma que lo editado ayer SE GUARDÓ (los endpoints actualizan todo el lote + commit).</div>'
        '<div class="conteo"><table><thead><tr><th>Fecha</th><th>Usuario</th><th>Acción</th><th>Detalle</th></tr></thead>'
        '<tbody>' + corr_filas + '</tbody></table></div>')
    # Lotes ACTUALMENTE en cuarentena (bloqueados para producción) · para que el
    # equipo vea de un vistazo qué falta liberar (Calidad) durante el cuadre.
    cuar_rows = []
    try:
        c.execute(
            "SELECT m.material_id, COALESCE(m.material_nombre,''), COALESCE(m.lote,''), "
            "COALESCE(m.cantidad,0), COALESCE(m.estanteria,''), COALESCE(m.estado_lote,''), "
            "COALESCE(m.fecha,'') FROM movimientos m "
            "WHERE UPPER(COALESCE(m.estado_lote,'')) IN ('CUARENTENA','CUARENTENA_EXTENDIDA') "
            "AND m.tipo='Entrada' ORDER BY m.fecha DESC LIMIT 300")
        cuar_rows = c.fetchall()
    except Exception:
        cuar_rows = []
    cuar_filas = ''.join(
        '<tr><td class="mono">' + _hh.escape(str(q[0])) + '</td>'
        '<td>' + _hh.escape(str(q[1])) + '</td>'
        '<td class="mono">' + _hh.escape(str(q[2])) + '</td>'
        '<td class="num">' + ('{:,.0f}'.format(q[3])) + '</td>'
        '<td>' + _hh.escape(str(q[4]) or '—') + '</td>'
        '<td><span style="color:#b45309;font-weight:700">' + _hh.escape(str(q[5])) + '</span></td>'
        '<td class="mono" style="font-size:11px">' + _hh.escape(str(q[6])[:16]) + '</td></tr>'
        for q in cuar_rows)
    if not cuar_filas:
        cuar_filas = '<tr><td colspan="7" class="muted">✓ No hay lotes bloqueados en cuarentena.</td></tr>'
    cuar_html = (
        '<h2 style="color:#b45309;margin:22px 0 8px;font-size:18px">🔒 Lotes en CUARENTENA (bloqueados para producción)</h2>'
        '<div class="sub" style="margin-bottom:10px">Estos lotes NO se pueden usar en producción hasta que <b>Calidad los libere</b> '
        '(Calidad → Cuarentena → Revisar CC). Al liberar, indica la ubicación final y salen del estante CUARENTENA.</div>'
        '<div class="conteo"><table><thead><tr><th>Código</th><th>Material</th><th>Lote</th><th class="num">Cantidad</th>'
        '<th>Estante físico</th><th>Estado</th><th>Ingreso</th></tr></thead><tbody>' + cuar_filas + '</tbody></table></div>')
    html = (
        '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1.0">'
        '<title>Rescate de Conteo · EOS</title><style>'
        '*{box-sizing:border-box;font-family:"Segoe UI",Roboto,sans-serif}'
        'body{margin:0;background:#f5f3ff;color:#0f172a;padding:20px}'
        '.wrap{max-width:1100px;margin:0 auto}'
        'h1{color:#7c3aed;margin:0 0 4px}.sub{color:#64748b;font-size:13px;margin-bottom:18px}'
        '.kpi{background:#fff;border-radius:12px;padding:14px 18px;margin-bottom:16px;box-shadow:0 2px 10px rgba(124,58,237,.08);font-size:14px}'
        '.kpi b{color:#16a34a;font-size:20px}'
        '.conteo{background:#fff;border-radius:12px;padding:14px 16px;margin-bottom:14px;box-shadow:0 2px 10px rgba(76,29,149,.07)}'
        '.ch{display:flex;justify-content:space-between;font-size:14px;margin-bottom:3px}'
        '.cm{font-size:11.5px;color:#94a3b8;margin-bottom:8px}'
        'table{width:100%;border-collapse:collapse;font-size:12px}'
        'th{background:#f5f3ff;color:#6d28d9;text-align:left;padding:7px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.3px}'
        'td{padding:6px 8px;border-bottom:1px solid #f1f5f9}'
        '.mono{font-family:ui-monospace,monospace;color:#1e40af}.num{text-align:right;font-variant-numeric:tabular-nums}.c{text-align:center}.muted{color:#94a3b8}'
        'a.bk{display:inline-block;background:#fff;color:#7c3aed;text-decoration:none;font-weight:700;font-size:13px;padding:8px 14px;border-radius:9px;border:1px solid #e9d5ff;margin-bottom:12px}'
        '</style></head><body><div class="wrap">'
        '<a class="bk" href="/inventarios">&larr; Volver</a>'
        '<h1>🛟 Rescate de Conteo Cíclico</h1>'
        '<div class="sub">Todos los conteos de los últimos ' + str(dias) + ' días, con sus items — sin filtros de fecha/semana/estado. '
        'Esto demuestra que lo ingresado NO se perdió. (cambia ?dias=N en la URL)</div>'
        '<div class="kpi">📦 <b>' + str(len(conteos)) + '</b> conteos · <b>' + str(tot_items) + '</b> items · <b>' + str(len(corr_rows)) + '</b> correcciones Bodega · <b style="color:#b45309">' + str(len(cuar_rows)) + '</b> lotes en cuarentena · últimos ' + str(dias) + ' días.</div>'
        + cuar_html
        + corr_html
        + '<h2 style="color:#7c3aed;margin:22px 0 8px;font-size:18px">📋 Conteos cíclicos</h2>'
        + cuerpo + '</div></body></html>')
    return Response(html, mimetype='text/html')

@bp.route('/api/admin/mee/diag', methods=['GET'])
def admin_mee_diag():
    """Diagnóstico completo de maestro_mee · detecta inconsistencias.
    Sebastián 24-may noche: 'Normalizar MEE sesión dedicada'.
    Devuelve JSON con: kpis + duplicados + descripciones vacías +
    espacios dobles + categorías mixtas + drift stock + huérfanos.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    # KPIs
    try:
        total_act = c.execute("SELECT COUNT(*) FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'").fetchone()[0]
    except Exception:
        total_act = 0
    try:
        total_arch = c.execute("SELECT COUNT(*) FROM maestro_mee WHERE COALESCE(estado,'Activo')!='Activo'").fetchone()[0]
    except Exception:
        total_arch = 0

    # Duplicados (case-insensitive + trim)
    duplicados = []
    try:
        for r in c.execute(
            """SELECT UPPER(TRIM(codigo)) AS k, COUNT(*) AS n,
                      GROUP_CONCAT(codigo, ' | ') AS lista
               FROM maestro_mee
               GROUP BY UPPER(TRIM(codigo))
               HAVING COUNT(*) > 1
               ORDER BY n DESC LIMIT 100"""
        ).fetchall():
            duplicados.append({'clave': r[0], 'n': r[1], 'codigos': r[2]})
    except Exception:
        pass

    # Descripciones vacías o con espacios dobles
    sin_desc = []
    desc_dobles = []
    try:
        for r in c.execute(
            "SELECT codigo, COALESCE(descripcion,'') FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'"
        ).fetchall():
            cod, desc = r[0], r[1] or ''
            if not desc.strip():
                sin_desc.append({'codigo': cod})
            elif '  ' in desc:
                desc_dobles.append({'codigo': cod, 'descripcion': desc})
    except Exception:
        pass

    # Categorías
    cats = {}
    try:
        for r in c.execute(
            "SELECT COALESCE(categoria,'(vacío)'), COUNT(*) FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo' GROUP BY categoria"
        ).fetchall():
            cats[r[0]] = r[1]
    except Exception:
        pass

    # Unidades
    unids = {}
    try:
        for r in c.execute(
            "SELECT COALESCE(unidad,'(vacío)'), COUNT(*) FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo' GROUP BY unidad"
        ).fetchall():
            unids[r[0]] = r[1]
    except Exception:
        pass

    # Stock drift
    drift_items = []
    try:
        # stock canónico = SUM movimientos_mee con UPPER(TRIM(tipo)) (fix mig 167)
        canon_stock = {}
        for r in c.execute(
            """SELECT mee_codigo,
                      COALESCE(SUM(CASE
                          WHEN UPPER(TRIM(tipo)) = 'ENTRADA' THEN ABS(cantidad)
                          WHEN UPPER(TRIM(tipo)) IN ('SALIDA','CONSUMO') THEN -ABS(cantidad)
                          WHEN UPPER(TRIM(tipo)) LIKE 'AJUSTE%' THEN cantidad
                          ELSE 0 END), 0)
               FROM movimientos_mee
               WHERE COALESCE(anulado,0) = 0
               GROUP BY mee_codigo"""
        ).fetchall():
            canon_stock[r[0]] = float(r[1] or 0)
        for r in c.execute(
            "SELECT codigo, descripcion, COALESCE(stock_actual,0) FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'"
        ).fetchall():
            cod = r[0]
            cache = float(r[2] or 0)
            real = canon_stock.get(cod, 0)
            delta = abs(real - cache)
            if delta > 1:  # tolerancia 1 unidad
                drift_items.append({
                    'codigo': cod, 'descripcion': r[1] or '',
                    'cache': round(cache, 2), 'real': round(real, 2),
                    'delta': round(delta, 2),
                })
    except Exception:
        pass
    drift_items.sort(key=lambda x: -x['delta'])

    # Huérfanos (sin uso en sku_mee_config ni produccion_checklist activo)
    huerfanos = []
    try:
        usados = set()
        try:
            for r in c.execute("SELECT DISTINCT UPPER(TRIM(mee_codigo)) FROM sku_mee_config WHERE COALESCE(aplica,1)=1").fetchall():
                if r[0]: usados.add(r[0])
        except Exception:
            pass
        try:
            for r in c.execute("SELECT DISTINCT UPPER(TRIM(mee_codigo_asignado)) FROM produccion_checklist WHERE COALESCE(mee_codigo_asignado,'') != ''").fetchall():
                if r[0]: usados.add(r[0])
        except Exception:
            pass
        for r in c.execute(
            "SELECT codigo, descripcion, COALESCE(stock_actual,0) FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'"
        ).fetchall():
            if (r[0] or '').upper().strip() not in usados:
                huerfanos.append({
                    'codigo': r[0], 'descripcion': r[1] or '',
                    'stock': float(r[2] or 0),
                })
    except Exception:
        pass

    return jsonify({
        'kpis': {
            'total_activos': total_act,
            'total_archivados': total_arch,
            'duplicados': len(duplicados),
            'sin_descripcion': len(sin_desc),
            'descripciones_dobles': len(desc_dobles),
            'categorias_distintas': len(cats),
            'unidades_distintas': len(unids),
            'drift_items': len(drift_items),
            'huerfanos': len(huerfanos),
        },
        'duplicados': duplicados[:50],
        'sin_descripcion': sin_desc[:50],
        'descripciones_dobles': desc_dobles[:50],
        'categorias': cats,
        'unidades': unids,
        'drift_items': drift_items[:50],
        'huerfanos': huerfanos[:100],
    })


@bp.route('/api/admin/mee/normalizar-descripciones', methods=['POST'])
def admin_mee_normalizar_descripciones():
    """Colapsa espacios dobles + TRIM en descripcion."""
    if 'compras_user' not in session or session.get('compras_user') not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    conn = get_db(); c = conn.cursor()
    user = session.get('compras_user','')
    n_fix = 0
    try:
        # Detectar items con dobles espacios o trim necesario
        for r in c.execute(
            "SELECT codigo, descripcion FROM maestro_mee WHERE descripcion IS NOT NULL"
        ).fetchall():
            cod, desc = r[0], r[1] or ''
            norm = ' '.join(desc.split())
            if norm != desc and norm:
                c.execute(
                    "UPDATE maestro_mee SET descripcion = ? WHERE codigo = ?",
                    (norm, cod),
                )
                n_fix += 1
        c.execute(
            """INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
               VALUES (?,?,?,?,?,?,datetime('now','-5 hours'))""",
            (user, 'MEE_NORMALIZAR_DESC', 'maestro_mee', 'bulk',
             f'Normalizadas {n_fix} descripciones', ''),
        )
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    return jsonify({'ok': True, 'normalizadas': n_fix})


@bp.route('/api/admin/mee/unificar-categorias', methods=['POST'])
def admin_mee_unificar_categorias():
    """Aplica mapping de categorías para unificar variantes."""
    if 'compras_user' not in session or session.get('compras_user') not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    body = request.get_json(silent=True) or {}
    mapping = body.get('mapping') or {}
    if not mapping:
        return jsonify({'error': 'mapping requerido'}), 400
    conn = get_db(); c = conn.cursor()
    user = session.get('compras_user','')
    n_fix = 0
    try:
        for de, hacia in mapping.items():
            de_s, h_s = (de or '').strip(), (hacia or '').strip()
            if not de_s or not h_s or de_s == h_s:
                continue
            res = c.execute(
                "UPDATE maestro_mee SET categoria = ? WHERE COALESCE(categoria,'') = ?",
                (h_s, de_s),
            )
            n_fix += res.rowcount or 0
        c.execute(
            """INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
               VALUES (?,?,?,?,?,?,datetime('now','-5 hours'))""",
            (user, 'MEE_UNIFICAR_CATEGORIAS', 'maestro_mee', 'bulk',
             f'Items recategorizados: {n_fix}', ''),
        )
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    return jsonify({'ok': True, 'recategorizados': n_fix})


@bp.route('/api/admin/mee/reconciliar-stock-bulk', methods=['POST'])
def admin_mee_reconciliar_stock_bulk():
    """Reconcilia stock_actual = SUM(movimientos_mee canónico) para
    todos los items con drift > tolerancia."""
    if 'compras_user' not in session or session.get('compras_user') not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    body = request.get_json(silent=True) or {}
    try:
        tol = float(body.get('tolerancia', 1.0))
    except (TypeError, ValueError):
        tol = 1.0
    conn = get_db(); c = conn.cursor()
    user = session.get('compras_user','')
    n_fix = 0
    items = []
    try:
        canon_stock = {}
        for r in c.execute(
            """SELECT mee_codigo,
                      COALESCE(SUM(CASE
                          WHEN UPPER(TRIM(tipo)) = 'ENTRADA' THEN ABS(cantidad)
                          WHEN UPPER(TRIM(tipo)) IN ('SALIDA','CONSUMO') THEN -ABS(cantidad)
                          WHEN UPPER(TRIM(tipo)) LIKE 'AJUSTE%' THEN cantidad
                          ELSE 0 END), 0)
               FROM movimientos_mee
               WHERE COALESCE(anulado,0) = 0
               GROUP BY mee_codigo"""
        ).fetchall():
            canon_stock[r[0]] = max(float(r[1] or 0), 0)
        for r in c.execute(
            "SELECT codigo, COALESCE(stock_actual,0) FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'"
        ).fetchall():
            cod = r[0]
            cache = float(r[1] or 0)
            real = canon_stock.get(cod, 0)
            if abs(real - cache) > tol:
                c.execute(
                    "UPDATE maestro_mee SET stock_actual = ? WHERE codigo = ?",
                    (real, cod),
                )
                items.append({'codigo': cod, 'cache': cache, 'real': real})
                n_fix += 1
        c.execute(
            """INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
               VALUES (?,?,?,?,?,?,datetime('now','-5 hours'))""",
            (user, 'MEE_RECONCILIAR_BULK', 'maestro_mee', 'bulk',
             f'Items reconciliados: {n_fix} (tol={tol})', ''),
        )
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    return jsonify({'ok': True, 'reconciliados': n_fix, 'items': items[:30]})


@bp.route('/api/admin/mee/marcar-huerfanos-inactivos', methods=['POST'])
def admin_mee_marcar_huerfanos_inactivos():
    """Marca como Inactivo los MEE que NO se usan en sku_mee_config
    ni produccion_checklist · respeta los que tengan stock > 0 si el
    body trae preservar_con_stock=1."""
    if 'compras_user' not in session or session.get('compras_user') not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    body = request.get_json(silent=True) or {}
    preservar_con_stock = bool(body.get('preservar_con_stock', True))
    conn = get_db(); c = conn.cursor()
    user = session.get('compras_user','')
    n_fix = 0
    try:
        usados = set()
        for r in c.execute("SELECT DISTINCT UPPER(TRIM(mee_codigo)) FROM sku_mee_config WHERE COALESCE(aplica,1)=1").fetchall():
            if r[0]: usados.add(r[0])
        for r in c.execute("SELECT DISTINCT UPPER(TRIM(mee_codigo_asignado)) FROM produccion_checklist WHERE COALESCE(mee_codigo_asignado,'') != ''").fetchall():
            if r[0]: usados.add(r[0])
        # FIX 27-may (P1) · stock CANÓNICO = SUM(movimientos_mee) · NO cache.
        # Antes: usaba maestro_mee.stock_actual (cache) → si drift estaba en 0
        # marcaba Inactivo MEEs que SÍ tenían stock real → invisibilidad de stock.
        for r in c.execute(
            "SELECT codigo FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'"
        ).fetchall():
            cod = r[0]
            if (cod or '').upper().strip() in usados:
                continue
            stock_real = 0.0
            if preservar_con_stock:
                # Stock real desde kardex MEE · Entrada=+, Salida=-, anulado excluído
                try:
                    s_row = c.execute(
                        """SELECT COALESCE(SUM(CASE
                              WHEN tipo='Entrada' THEN cantidad
                              WHEN tipo='Salida'  THEN -cantidad
                              WHEN tipo='Ajuste'  THEN cantidad
                              ELSE 0 END),0)
                           FROM movimientos_mee
                           WHERE mee_codigo=? AND COALESCE(anulado,0)=0""",
                        (cod,)
                    ).fetchone()
                    stock_real = float((s_row or [0])[0] or 0)
                except Exception:
                    stock_real = 0.0
                if stock_real > 0:
                    continue
            c.execute(
                "UPDATE maestro_mee SET estado = 'Inactivo' WHERE codigo = ?",
                (cod,),
            )
            n_fix += 1
        c.execute(
            """INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
               VALUES (?,?,?,?,?,?,datetime('now','-5 hours'))""",
            (user, 'MEE_MARCAR_HUERFANOS', 'maestro_mee', 'bulk',
             f'Items archivados: {n_fix} (preservar_stock={preservar_con_stock})', ''),
        )
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    return jsonify({'ok': True, 'archivados': n_fix})


@bp.route('/admin/normalizar-mee', methods=['GET'])
def admin_normalizar_mee_pagina():
    """Página HTML standalone para auditar + normalizar maestro_mee.
    Dashboard + tabs por tipo de problema + botones fix masivo.
    """
    if 'compras_user' not in session:
        return "<html><body><h2>No autorizado</h2></body></html>", 401
    user = session.get('compras_user','')
    if user not in ADMIN_USERS:
        return "<html><body><h2>Solo admin</h2></body></html>", 403
    return """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Normalizar MEE</title>
<style>
  body{font-family:system-ui,-apple-system,Arial;background:#f8fafc;margin:0;padding:24px}
  .card{max-width:1400px;margin:0 auto 14px;background:#fff;border-radius:14px;padding:24px;box-shadow:0 4px 20px rgba(0,0,0,.08)}
  h1{color:#1e293b;margin:0 0 8px;font-size:22px}
  h2{color:#7c3aed;margin:18px 0 12px;font-size:16px;border-bottom:2px solid #7c3aed;padding-bottom:6px}
  .kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin:14px 0}
  .kpi{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px;text-align:center}
  .kpi-val{font-size:24px;font-weight:800;color:#1e293b}
  .kpi-lbl{font-size:11px;color:#64748b;text-transform:uppercase;margin-top:2px}
  .kpi.warn{background:#fef3c7;border-color:#f59e0b}
  .kpi.warn .kpi-val{color:#92400e}
  .kpi.bad{background:#fee2e2;border-color:#dc2626}
  .kpi.bad .kpi-val{color:#991b1b}
  table{width:100%;border-collapse:collapse;font-size:12px;margin-top:8px}
  th{background:#f1f5f9;padding:8px;text-align:left;font-size:11px;text-transform:uppercase;color:#475569}
  td{border-bottom:1px solid #f1f5f9;padding:8px}
  .btn{padding:8px 14px;border-radius:6px;border:none;font-size:12px;font-weight:700;cursor:pointer;margin-right:6px}
  .btn-prim{background:#7c3aed;color:#fff}
  .btn-warn{background:#ea580c;color:#fff}
  .btn-danger{background:#dc2626;color:#fff}
  .btn-link{background:transparent;color:#475569;border:1px solid #e2e8f0}
  .tab{display:none}
  .tab.active{display:block}
  .tab-btns{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px}
  .tab-btn{padding:8px 14px;background:#f1f5f9;color:#475569;border:1px solid #e2e8f0;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600}
  .tab-btn.active{background:#7c3aed;color:#fff;border-color:#7c3aed}
  input,select{padding:6px 10px;border:1px solid #cbd5e1;border-radius:5px;font-size:12px}
  #msg{margin:10px 0;font-size:12px}
</style></head><body>

<div class="card">
  <h1>🧹 Normalizar maestro_mee</h1>
  <p style="color:#475569;font-size:13px">Limpieza + estandarización de Materiales de Envase. Cada acción es auditada en audit_log.</p>
  <button class="btn btn-prim" onclick="cargarDiag()">🔍 Diagnosticar</button>
  <a href="/admin/mee-fugas-check" class="btn btn-link" style="text-decoration:none;line-height:1.4">⚠ Mee fugas</a>
  <a href="/" class="btn btn-link" style="text-decoration:none;line-height:1.4">← Dashboard</a>
  <div id="msg"></div>
</div>

<div class="card" id="card-kpis" style="display:none">
  <h2>📊 Estado actual</h2>
  <div class="kpi-grid" id="kpis"></div>
</div>

<div class="card" id="card-tabs" style="display:none">
  <div class="tab-btns">
    <button class="tab-btn active" data-tab="dup" onclick="setTab('dup')">🔁 Duplicados</button>
    <button class="tab-btn" data-tab="desc" onclick="setTab('desc')">📝 Descripciones</button>
    <button class="tab-btn" data-tab="cat" onclick="setTab('cat')">🏷 Categorías</button>
    <button class="tab-btn" data-tab="drift" onclick="setTab('drift')">⚖ Stock drift</button>
    <button class="tab-btn" data-tab="huer" onclick="setTab('huer')">🍃 Huérfanos</button>
  </div>

  <div class="tab active" id="t-dup">
    <h2>🔁 Códigos duplicados (case-insensitive + trim)</h2>
    <p style="color:#475569;font-size:12px">Items con códigos similares (e.g. 'ENV-100' vs 'env-100'). Fix manual sugerido · revisar uno por uno antes de merge.</p>
    <div id="tbl-dup"></div>
  </div>

  <div class="tab" id="t-desc">
    <h2>📝 Descripciones inconsistentes</h2>
    <p style="color:#475569;font-size:12px">Espacios dobles + vacías. Fix masivo: colapsa espacios y aplica TRIM.</p>
    <button class="btn btn-warn" onclick="normalizarDescripciones()">🔧 Normalizar descripciones (TRIM + colapsar espacios)</button>
    <div id="tbl-desc"></div>
  </div>

  <div class="tab" id="t-cat">
    <h2>🏷 Categorías</h2>
    <p style="color:#475569;font-size:12px">Distribución actual de categorías. Si hay variantes ('Envase' vs 'envase' vs 'ENVASE'), unificá con el mapping.</p>
    <div id="tbl-cat"></div>
    <div style="margin-top:18px;border-top:1px solid #e2e8f0;padding-top:14px">
      <h3 style="margin:0 0 8px;color:#1e293b;font-size:14px">Aplicar mapping de unificación</h3>
      <div id="mapping-rows"></div>
      <button class="btn btn-link" onclick="addMappingRow()">+ Agregar regla</button>
      <button class="btn btn-warn" onclick="aplicarMapping()">🔧 Aplicar mapping</button>
    </div>
  </div>

  <div class="tab" id="t-drift">
    <h2>⚖ Stock drift (cache vs movimientos_mee)</h2>
    <p style="color:#475569;font-size:12px">stock_actual de maestro_mee desincronizado con SUM canónico de movimientos.</p>
    <div style="margin-bottom:8px"><label>Tolerancia: <input id="tol" type="number" value="1" step="0.1" min="0" style="width:80px"></label>
    <button class="btn btn-warn" onclick="reconciliarBulk()">🔧 Reconciliar todos (bulk)</button></div>
    <div id="tbl-drift"></div>
  </div>

  <div class="tab" id="t-huer">
    <h2>🍃 Items huérfanos (sin uso en sku_mee_config ni produccion_checklist)</h2>
    <p style="color:#475569;font-size:12px">MEEs activos que ningún producto ni checklist usa.</p>
    <label style="display:block;margin:8px 0"><input type="checkbox" id="preservar" checked> Preservar items con stock_actual &gt; 0</label>
    <button class="btn btn-danger" onclick="archivarHuerfanos()">🗄 Archivar huérfanos (Inactivo)</button>
    <div id="tbl-huer"></div>
  </div>
</div>

<script>
let DATA=null;
let MAPPING_N=0;

function msg(text, ok){
  const el=document.getElementById('msg');
  el.innerHTML='<span style="color:'+(ok?'#15803d':'#dc2626')+';font-weight:700">'+text+'</span>';
  setTimeout(()=>{el.innerHTML=''},5000);
}

async function cargarDiag(){
  msg('Calculando…', true);
  try{
    const r=await fetch('/api/admin/mee/diag');
    const d=await r.json();
    if(!r.ok){msg('Error: '+(d.error||r.status));return}
    DATA=d;
    renderKPIs();
    renderTabs();
    document.getElementById('card-kpis').style.display='block';
    document.getElementById('card-tabs').style.display='block';
    msg('✓ Diagnóstico completo', true);
  }catch(e){msg('Error red: '+e.message)}
}

function renderKPIs(){
  const k=DATA.kpis;
  const html=[
    ['Items activos', k.total_activos, ''],
    ['Archivados', k.total_archivados, ''],
    ['Duplicados', k.duplicados, k.duplicados>0?'warn':''],
    ['Sin descripción', k.sin_descripcion, k.sin_descripcion>0?'warn':''],
    ['Espacios dobles', k.descripciones_dobles, k.descripciones_dobles>0?'warn':''],
    ['Categorías distintas', k.categorias_distintas, k.categorias_distintas>5?'warn':''],
    ['Unidades distintas', k.unidades_distintas, k.unidades_distintas>3?'warn':''],
    ['Stock drift', k.drift_items, k.drift_items>0?'bad':''],
    ['Huérfanos', k.huerfanos, k.huerfanos>0?'warn':''],
  ].map(([l,v,c])=>'<div class="kpi '+c+'"><div class="kpi-val">'+v+'</div><div class="kpi-lbl">'+l+'</div></div>').join('');
  document.getElementById('kpis').innerHTML=html;
}

function renderTabs(){
  // Duplicados
  document.getElementById('tbl-dup').innerHTML = DATA.duplicados.length===0
    ? '<p style="color:#15803d">✓ Ningún duplicado detectado</p>'
    : '<table><thead><tr><th>Clave normalizada</th><th>N</th><th>Códigos reales</th></tr></thead><tbody>'
      + DATA.duplicados.map(x=>'<tr><td><code>'+x.clave+'</code></td><td>'+x.n+'</td><td style="font-size:11px">'+x.codigos+'</td></tr>').join('')
      + '</tbody></table>';
  // Descripciones
  const dDesc = (DATA.sin_descripcion.length+DATA.descripciones_dobles.length);
  if(!dDesc){
    document.getElementById('tbl-desc').innerHTML='<p style="color:#15803d">✓ Todas las descripciones limpias</p>';
  }else{
    let h='';
    if(DATA.sin_descripcion.length){
      h+='<h3 style="margin:14px 0 6px;font-size:13px">Sin descripción ('+DATA.sin_descripcion.length+')</h3>';
      h+='<table><thead><tr><th>Código</th></tr></thead><tbody>'+DATA.sin_descripcion.map(x=>'<tr><td><code>'+x.codigo+'</code></td></tr>').join('')+'</tbody></table>';
    }
    if(DATA.descripciones_dobles.length){
      h+='<h3 style="margin:14px 0 6px;font-size:13px">Con dobles espacios ('+DATA.descripciones_dobles.length+')</h3>';
      h+='<table><thead><tr><th>Código</th><th>Descripción actual</th></tr></thead><tbody>'+DATA.descripciones_dobles.map(x=>'<tr><td><code>'+x.codigo+'</code></td><td>'+x.descripcion.replace(/ /g,'·')+'</td></tr>').join('')+'</tbody></table>';
    }
    document.getElementById('tbl-desc').innerHTML=h;
  }
  // Categorías
  const cats=DATA.categorias;
  document.getElementById('tbl-cat').innerHTML='<table><thead><tr><th>Categoría</th><th>Count</th></tr></thead><tbody>'
    + Object.entries(cats).sort((a,b)=>b[1]-a[1]).map(([k,v])=>'<tr><td><code>'+k+'</code></td><td>'+v+'</td></tr>').join('')
    + '</tbody></table>';
  // Drift
  document.getElementById('tbl-drift').innerHTML = DATA.drift_items.length===0
    ? '<p style="color:#15803d">✓ Sin drift detectado</p>'
    : '<table><thead><tr><th>Código</th><th>Descripción</th><th style="text-align:right">Cache</th><th style="text-align:right">Real (SUM)</th><th style="text-align:right">Delta</th></tr></thead><tbody>'
      + DATA.drift_items.map(x=>'<tr><td><code>'+x.codigo+'</code></td><td>'+x.descripcion+'</td><td style="text-align:right">'+x.cache.toLocaleString()+'</td><td style="text-align:right">'+x.real.toLocaleString()+'</td><td style="text-align:right;font-weight:700;color:#dc2626">'+x.delta.toLocaleString()+'</td></tr>').join('')
      + '</tbody></table>';
  // Huérfanos
  document.getElementById('tbl-huer').innerHTML = DATA.huerfanos.length===0
    ? '<p style="color:#15803d">✓ Sin huérfanos</p>'
    : '<table><thead><tr><th>Código</th><th>Descripción</th><th style="text-align:right">Stock</th></tr></thead><tbody>'
      + DATA.huerfanos.map(x=>'<tr><td><code>'+x.codigo+'</code></td><td>'+x.descripcion+'</td><td style="text-align:right">'+x.stock.toLocaleString()+'</td></tr>').join('')
      + '</tbody></table>';
}

function setTab(tab){
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('t-'+tab).classList.add('active');
  document.querySelector('[data-tab='+tab+']').classList.add('active');
}

async function normalizarDescripciones(){
  if(!confirm('¿Aplicar TRIM + colapsar espacios dobles en todas las descripciones?')) return;
  const r=await fetch('/api/admin/mee/normalizar-descripciones',{method:'POST',headers:{'Content-Type':'application/json'}});
  const d=await r.json();
  if(!r.ok){msg('Error: '+(d.error||r.status));return}
  msg('✓ '+d.normalizadas+' descripciones normalizadas',true);
  cargarDiag();
}

function addMappingRow(){
  MAPPING_N++;
  const n=MAPPING_N;
  document.getElementById('mapping-rows').insertAdjacentHTML('beforeend',
    '<div style="display:flex;gap:6px;align-items:center;margin-bottom:4px" id="m-'+n+'">'
    +'<input id="m-de-'+n+'" placeholder="De: envase" style="flex:1">'
    +'<span>→</span>'
    +'<input id="m-h-'+n+'" placeholder="Hacia: Envase" style="flex:1">'
    +'<button class="btn btn-link" onclick="document.getElementById(\\''+'m-'+n+'\\').remove()">✕</button>'
    +'</div>');
}

async function aplicarMapping(){
  const mapping={};
  for(let i=1;i<=MAPPING_N;i++){
    const de=document.getElementById('m-de-'+i);
    const h=document.getElementById('m-h-'+i);
    if(de && h && de.value.trim() && h.value.trim()){
      mapping[de.value.trim()]=h.value.trim();
    }
  }
  if(Object.keys(mapping).length===0){msg('Agregá al menos 1 regla');return}
  if(!confirm('¿Aplicar mapping a categorías?\\n\\n'+JSON.stringify(mapping,null,2))) return;
  const r=await fetch('/api/admin/mee/unificar-categorias',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mapping})});
  const d=await r.json();
  if(!r.ok){msg('Error: '+(d.error||r.status));return}
  msg('✓ '+d.recategorizados+' items recategorizados',true);
  cargarDiag();
}

async function reconciliarBulk(){
  const tol=parseFloat(document.getElementById('tol').value)||1;
  if(!confirm('¿Reconciliar stock_actual de TODOS los MEE con drift > '+tol+'?')) return;
  const r=await fetch('/api/admin/mee/reconciliar-stock-bulk',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tolerancia:tol})});
  const d=await r.json();
  if(!r.ok){msg('Error: '+(d.error||r.status));return}
  msg('✓ '+d.reconciliados+' items reconciliados',true);
  cargarDiag();
}

async function archivarHuerfanos(){
  const preservar=document.getElementById('preservar').checked;
  if(!confirm('¿Archivar (activo=0) los huérfanos?'+(preservar?' Los con stock>0 se preservan.':' INCLUYE los con stock>0.'))) return;
  const r=await fetch('/api/admin/mee/marcar-huerfanos-inactivos',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({preservar_con_stock:preservar})});
  const d=await r.json();
  if(!r.ok){msg('Error: '+(d.error||r.status));return}
  msg('✓ '+d.archivados+' archivados',true);
  cargarDiag();
}

cargarDiag();
</script>
</body></html>"""


@bp.route('/planta/conteo-ciclico', methods=['GET'])
def planta_conteo_ciclico_pagina():
    """Sebastián 24-may noche · UI inventario cíclico responsive (tablet
    + móvil). El operario entra en la bodega con dispositivo móvil,
    elige estantería, ve la lista de MPs con stock_sistema, ingresa
    el stock_físico, guarda. Cierre opcional aplica ajustes <5% auto.

    Mobile-first · inputs grandes · sticky header · auto-save por blur.
    Reusa endpoints existentes /api/conteo/*.
    """
    if 'compras_user' not in session:
        # Branding fix 27-may · página de no-auth ahora consistente con EOS
        return ("""<!DOCTYPE html><html lang="es"><head>
<meta charset="utf-8"><title>Sesión requerida · EOS</title>
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5.0,user-scalable=yes">
<link rel="stylesheet" href="/static/cortex.css">
<style>body{display:flex;align-items:center;justify-content:center;min-height:100vh;background:var(--cx-bg,#0f172a);color:var(--cx-text,#f1f5f9);font-family:'Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif}
.box{background:var(--cx-card,#1e293b);border:1px solid var(--cx-border,#334155);border-radius:14px;padding:30px 36px;max-width:440px;text-align:center}
.brand{color:#a78bfa;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.15em;margin-bottom:12px}
h1{margin:0 0 12px;font-size:20px;color:#fff}
.btn{display:inline-block;margin-top:18px;padding:10px 22px;background:#7c3aed;color:#fff;border-radius:8px;text-decoration:none;font-weight:700}</style>
</head><body>
<div class="box">
  <div class="brand">EOS · HHA Group</div>
  <h1>Sesión requerida</h1>
  <p style="color:#94a3b8">Iniciá sesión para acceder al conteo cíclico.</p>
  <a class="btn" href="/">Ir al login</a>
</div></body></html>""", 401)
    return """<!DOCTYPE html>
<html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5.0,user-scalable=yes">
<title>Conteo Cíclico · Bodega</title>
<style>
  *{box-sizing:border-box}
  body{font-family:system-ui,-apple-system,Arial;background:#f1f5f9;margin:0;padding:0;font-size:15px}
  .top{position:sticky;top:0;background:#0f766e;color:#fff;padding:12px 16px;z-index:100;box-shadow:0 2px 8px rgba(0,0,0,.15)}
  .top h1{margin:0;font-size:17px;font-weight:700}
  .top .sub{font-size:11px;opacity:.9;margin-top:2px}
  .container{padding:14px;max-width:900px;margin:0 auto}
  .card{background:#fff;border-radius:10px;padding:14px;margin-bottom:12px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
  .btn{display:inline-block;padding:14px 24px;border-radius:10px;border:none;font-size:15px;font-weight:700;cursor:pointer;text-align:center;text-decoration:none;width:100%}
  .btn-prim{background:#0f766e;color:#fff}
  .btn-warn{background:#ea580c;color:#fff}
  .btn-danger{background:#dc2626;color:#fff}
  .btn-sec{background:#e2e8f0;color:#475569}
  label{display:block;font-size:11px;color:#64748b;margin-bottom:4px;font-weight:700;text-transform:uppercase}
  select,input{width:100%;padding:14px 12px;font-size:16px;border:2px solid #cbd5e1;border-radius:8px;font-family:inherit}
  input[type=number]{font-family:ui-monospace,monospace;text-align:right;font-size:18px;font-weight:700}
  .mp-row{padding:12px;border:1px solid #e2e8f0;border-radius:10px;margin-bottom:10px;background:#fff;transition:background .2s}
  .mp-row.dirty{background:#fef9c3;border-color:#ca8a04}
  .mp-row.saved{background:#dcfce7;border-color:#16a34a}
  .mp-row.diff-big{background:#fee2e2;border-color:#dc2626}
  .mp-info{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:10px;flex-wrap:wrap;gap:6px}
  .mp-cod{font-family:ui-monospace,monospace;font-weight:800;color:#1e293b;font-size:13px}
  .mp-nom{font-size:14px;color:#1e293b;font-weight:600;margin-top:2px}
  .mp-meta{font-size:11px;color:#64748b;font-family:ui-monospace,monospace}
  .mp-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px}
  .stock-box{background:#f8fafc;padding:10px;border-radius:8px;text-align:center}
  .stock-box .val{font-size:22px;font-weight:800;color:#1e293b;font-family:ui-monospace,monospace}
  .stock-box .lbl{font-size:10px;color:#64748b;text-transform:uppercase;margin-top:2px}
  .diff{font-weight:800;font-family:ui-monospace,monospace}
  .diff.pos{color:#15803d} .diff.neg{color:#991b1b}
  .progress{position:fixed;bottom:0;left:0;right:0;background:#fff;border-top:2px solid #e2e8f0;padding:12px 16px;z-index:90;display:flex;gap:8px}
  .progress .info{flex:1;font-size:12px;color:#475569}
  .progress strong{color:#0f766e;font-size:16px}
  .toast{position:fixed;top:60px;left:50%;transform:translateX(-50%);background:#15803d;color:#fff;padding:10px 20px;border-radius:8px;font-weight:700;z-index:200;display:none}
  .toast.err{background:#dc2626}
  @media(max-width:480px){.mp-grid{grid-template-columns:1fr}}
</style></head>
<body>
<div class="top">
  <h1>📦 Conteo Cíclico de Bodega</h1>
  <div class="sub" id="top-sub">Selecciona estantería para empezar</div>
</div>

<div id="toast" class="toast"></div>

<div class="container" id="vista-inicio">
  <div class="card">
    <label>Estantería a contar</label>
    <select id="sel-est"><option value="">— elegir —</option></select>
  </div>
  <div class="card">
    <label>Tu nombre (responsable)</label>
    <input id="responsable" type="text" placeholder="Tu nombre">
  </div>
  <button class="btn btn-prim" onclick="iniciarConteo()">▶ Iniciar conteo</button>
  <div style="margin-top:14px"><a href="/" class="btn btn-sec">← Volver al dashboard</a></div>
</div>

<div class="container" id="vista-conteo" style="display:none;padding-bottom:80px">
  <div class="card" style="background:#fef3c7;border-left:5px solid #ca8a04">
    <strong>Conteo #<span id="cnt-num"></span></strong> · Estantería <strong id="cnt-est"></strong><br>
    <span style="font-size:11px;color:#64748b">Stock sistema = lo que dice el kardex · Stock físico = lo que CONTAS en bodega · diferencia automática</span>
  </div>
  <input type="text" id="filtro" placeholder="🔍 Filtrar por código o nombre…" oninput="filtrar()" style="margin-bottom:10px">
  <div id="lista-mps"></div>
</div>

<div class="progress" id="progress" style="display:none">
  <div class="info"><strong id="prog-cont">0</strong> de <strong id="prog-total">0</strong> contados · <strong id="prog-diff" style="color:#dc2626">0</strong> con diferencia</div>
  <button class="btn btn-prim" style="padding:10px 16px;width:auto" onclick="guardarTodo()">💾 Guardar</button>
  <button class="btn btn-danger" style="padding:10px 16px;width:auto" onclick="cerrarConteo()">✓ Cerrar</button>
</div>

<script>
let CONTEO_ID=null, CONTEO_NUM='', EST='';
let ITEMS=[]; // {codigo_mp, nombre, stock_sistema, lote, ...}
let STOCK_FISICO={}; // codigo_lote -> valor

function toast(msg, err){
  const t=document.getElementById('toast');
  t.textContent=msg; t.className='toast'+(err?' err':'');
  t.style.display='block'; setTimeout(()=>{t.style.display='none'},3000);
}

async function cargarEstanterias(){
  try{
    const r=await fetch('/api/conteo/estanterias');
    const d=await r.json();
    const sel=document.getElementById('sel-est');
    (d.items||d||[]).forEach(e=>{
      const o=document.createElement('option');
      o.value=typeof e==='string'?e:e.estanteria;
      o.textContent=(typeof e==='string'?e:e.estanteria)+(e.n_mps?' ('+e.n_mps+' MPs)':'');
      sel.appendChild(o);
    });
  }catch(e){toast('Error cargando estanterías',true)}
}

async function iniciarConteo(){
  EST=document.getElementById('sel-est').value;
  const resp=document.getElementById('responsable').value.trim();
  if(!EST){toast('Elegí estantería',true);return}
  if(!resp){toast('Poné tu nombre',true);return}
  try{
    const r=await fetch('/api/conteo/iniciar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({estanteria:EST,responsable:resp})});
    const d=await r.json();
    if(!r.ok){toast('Error: '+(d.error||r.status),true);return}
    CONTEO_ID=d.conteo_id; CONTEO_NUM=d.numero;
    document.getElementById('cnt-num').textContent=d.numero;
    document.getElementById('cnt-est').textContent=EST;
    document.getElementById('top-sub').textContent=d.numero+' · '+EST+(d.resuming?' (retomado)':' (nuevo)');
    await cargarMPs();
    document.getElementById('vista-inicio').style.display='none';
    document.getElementById('vista-conteo').style.display='block';
    document.getElementById('progress').style.display='flex';
  }catch(e){toast('Error red: '+e.message,true)}
}

async function cargarMPs(){
  const r=await fetch('/api/conteo/materiales?estanteria='+encodeURIComponent(EST));
  const d=await r.json();
  // A-2 (Sebastian 12-jun): /api/conteo/materiales devuelve un ARRAY plano, no
  // {items:[...]}. Antes ITEMS=d.items||[] daba SIEMPRE [] -> la pagina movil de
  // conteo nunca listaba materiales al operario en bodega. Soporta ambos contratos.
  ITEMS = Array.isArray(d) ? d : (d.items||[]);
  // Si el conteo se está retomando, recuperar valores guardados previamente
  try{
    const r2=await fetch('/api/conteo/'+CONTEO_ID+'/items');
    const d2=await r2.json();
    // FIX 6-jun-2026 (incidente "lo de ayer no sale"): el endpoint devuelve un
    // array plano; antes se leía d2.items (undefined) → forEach nunca corría →
    // las casillas de stock físico salían en blanco al retomar el conteo de ayer.
    // Tolerante a ambas formas (array o {items:[...]}).
    const arr2 = Array.isArray(d2) ? d2 : (d2.items||[]);
    arr2.forEach(it=>{
      if(it.stock_fisico!=null && it.stock_fisico!==''){
        const k=it.codigo_mp+'|'+(it.lote||'');
        STOCK_FISICO[k]=it.stock_fisico;
      }
    });
  }catch(e){ console.error('No se pudieron recuperar los valores del conteo previo (retomar):', e); }
  renderItems();
}

function renderItems(){
  const filtro=(document.getElementById('filtro').value||'').toLowerCase().trim();
  const cont=document.getElementById('lista-mps');
  cont.innerHTML='';
  let contados=0, diffs=0;
  ITEMS.forEach((it,idx)=>{
    const k=it.codigo_mp+'|'+(it.lote||'');
    const fis=STOCK_FISICO[k];
    if(filtro){
      const blob=(it.codigo_mp+' '+(it.nombre||'')+' '+(it.lote||'')).toLowerCase();
      if(blob.indexOf(filtro)<0)return;
    }
    const diff=(fis!=null && fis!=='')?(fis - it.stock_sistema):null;
    const pctDiff=(diff!=null && it.stock_sistema>0)?Math.abs(diff/it.stock_sistema)*100:0;
    const cls = (fis==null||fis==='')?'':((pctDiff>5)?'mp-row diff-big':'mp-row saved');
    if(fis!=null && fis!=='') contados++;
    if(diff!=null && Math.abs(diff)>0.1) diffs++;
    const diffHtml = diff===null?'<span style="color:#94a3b8">—</span>':
      ('<span class="diff '+(diff>=0?'pos':'neg')+'">'+(diff>=0?'+':'')+diff.toFixed(1)+(pctDiff>5?' ⚠ '+pctDiff.toFixed(1)+'%':'')+'</span>');
    cont.insertAdjacentHTML('beforeend',
      '<div class="'+(cls||'mp-row')+'" data-idx="'+idx+'">'
      +'<div class="mp-info"><div><div class="mp-cod">'+esc(it.codigo_mp)+'</div><div class="mp-nom">'+esc(it.nombre||'')+'</div>'
      +'<div class="mp-meta">Lote: '+esc(it.lote||'—')+(it.fecha_vencimiento?' · vence '+it.fecha_vencimiento.slice(0,10):'')+'</div></div></div>'
      +'<div class="mp-grid">'
      +'<div class="stock-box"><div class="val">'+formatNum(it.stock_sistema)+'</div><div class="lbl">Sistema (g)</div></div>'
      +'<div><label>Stock físico (g)</label><input type="number" inputmode="decimal" min="0" step="0.01" value="'+(fis!=null?fis:'')+'" placeholder="contá y escribí" onchange="setFisico('+idx+',this.value)" onfocus="this.select()"></div>'
      +'</div>'
      +'<div style="margin-top:8px;text-align:center;font-size:13px">Diferencia: '+diffHtml+'</div>'
      +'</div>');
  });
  document.getElementById('prog-cont').textContent=contados;
  document.getElementById('prog-total').textContent=ITEMS.length;
  document.getElementById('prog-diff').textContent=diffs;
}

function setFisico(idx,val){
  const it=ITEMS[idx];
  const k=it.codigo_mp+'|'+(it.lote||'');
  if(val===''||val==null){delete STOCK_FISICO[k]; renderItems(); return}
  const n=parseFloat(val);
  if(isNaN(n)||n<0){toast('Valor inválido',true);return}
  STOCK_FISICO[k]=n;
  renderItems();
}

function filtrar(){renderItems()}

async function guardarTodo(){
  if(!CONTEO_ID)return;
  const payload={items:ITEMS.map(it=>{
    const k=it.codigo_mp+'|'+(it.lote||'');
    const fis=STOCK_FISICO[k];
    return {
      codigo_mp:it.codigo_mp, nombre:it.nombre, lote:it.lote||'',
      stock_sistema:it.stock_sistema, stock_fisico:(fis!=null?fis:''),
      estanteria:EST, precio_ref:it.precio_ref||0,
    };
  })};
  try{
    const r=await fetch('/api/conteo/'+CONTEO_ID+'/guardar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    const d=await r.json();
    if(!r.ok){toast('Error: '+(d.error||r.status),true);return}
    if((d.items_invalidos||[]).length){
      toast('⚠ '+d.items_invalidos.length+' items inválidos',true);
    }else{
      toast('✓ Guardado · '+d.items_con_diferencia+' con diferencia');
    }
  }catch(e){toast('Error red: '+e.message,true)}
}

async function cerrarConteo(){
  if(!CONTEO_ID)return;
  await guardarTodo();
  if(!confirm('¿Cerrar el conteo?\\n\\nTodas las diferencias se ajustan en el kardex. Las grandes quedan en el informe de revisión (sin frenar el cierre).')) return;
  try{
    const r=await fetch('/api/conteo/'+CONTEO_ID+'/cerrar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({})});
    const d=await r.json();
    if(!r.ok){toast('Error: '+(d.error||r.status),true);return}
    // B-2 (12-jun): el backend devuelve total_items_ajustados/total_pendientes
    // (antes leia ajustes_aplicados/items_gerencia -> "undefined").
    const msg='Cerrado · '+(d.total_items_ajustados||0)+' ajuste(s) aplicado(s) al kardex';
    alert(msg+'\\n\\nVolvé al dashboard.');
    location.href='/';
  }catch(e){toast('Error red: '+e.message,true)}
}

function esc(s){const d=document.createElement('div');d.textContent=s||'';return d.innerHTML}
function formatNum(n){return (parseFloat(n)||0).toLocaleString('es-CO',{maximumFractionDigits:1})}

cargarEstanterias();
</script>
</body></html>"""


@bp.route('/api/conteo/<int:conteo_id>/items', methods=['GET'])
def conteo_get_items(conteo_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM conteo_items WHERE conteo_id=? ORDER BY codigo_mp", (conteo_id,))
    rows = c.fetchall()
    cols = [d[0] for d in c.description]
    # 6-jun-2026 · devolver {items:[...]} (lo que espera el frontend al retomar).
    # El frontend también tolera array plano, pero unificamos el contrato.
    return jsonify({'items': [dict(zip(cols, r)) for r in rows]})

@bp.route('/api/lotes/cc-review', methods=['POST'])
def cc_review():
    # Antes era hardcoded {hernando} + ADMIN_USERS. Ahora usa CALIDAD_USERS
    # (laura, miguel, yuliel) + ADMIN — consistente con la matriz de permisos.
    user, err, code = _require_qc()
    if err:
        return err, code
    d = request.json or {}
    mov_id = d.get('mov_id')
    if not mov_id:
        return jsonify({'error': 'mov_id requerido'}), 400
    solubilidad = d.get('solubilidad', '')
    resultado_aql = d.get('resultado_aql', '')
    if solubilidad == 'RECHAZO' or resultado_aql == 'NO_CONFORME':
        estado_final = 'RECHAZADO'
    elif resultado_aql == 'CUARENTENA_EXTENDIDA':
        estado_final = 'CUARENTENA_EXTENDIDA'
    else:
        estado_final = 'APROBADO'
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id, material_id, lote, estado_lote FROM movimientos WHERE id=?", (mov_id,))
    mov = c.fetchone()
    if not mov:
        return jsonify({'error': 'Lote no encontrado'}), 404
    if mov[3] not in ('CUARENTENA', 'CUARENTENA_EXTENDIDA'):
        return jsonify({'error': 'Lote no esta en cuarentena'}), 400
    # Firma electrónica Part 11 §11.200 · CC review dispone un lote regulado
    # (aprobar/rechazar/cuarentena extendida) · exige re-autenticación vía
    # /api/sign. El meaning depende de la disposición resultante.
    meaning = ('rechaza' if estado_final == 'RECHAZADO'
               else 'libera' if estado_final == 'APROBADO'
               else 'aprueba')  # CUARENTENA_EXTENDIDA
    sig_id = d.get('signature_id')
    if not _validar_e_sign(c, sig_id, record_table='movimientos', record_id=mov_id,
                           meaning=meaning, signer_username=user):
        return jsonify({
            'error': 'Firma electrónica requerida',
            'requiere_firma': True,
            'sign_meaning': meaning,
            'record_table': 'movimientos',
            'record_id': str(mov_id),
            'estado_propuesto': estado_final,
            'detail': f"Firmá vía POST /api/sign con meaning='{meaning}', "
                      f"record_table='movimientos', record_id='{mov_id}' y reenviá signature_id.",
        }), 400
    c.execute(
        "INSERT INTO cc_reviews (mov_id,lote,codigo_mp,coa_ok,lote_coincide,coa_vigente,ficha_ok,"
        "solubilidad,resultado_aql,observaciones_aql,muestra_retencion,observaciones,firmante,estado_final,fecha,ip) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now', '-5 hours'),?)",
        (mov_id, d.get('lote',''), d.get('codigo_mp',''),
         1 if d.get('coa_ok') else 0, 1 if d.get('lote_coincide') else 0,
         1 if d.get('coa_vigente') else 0, 1 if d.get('ficha_ok') else 0,
         solubilidad, resultado_aql, d.get('observaciones_aql',''),
         1 if d.get('muestra_retencion') else 0, d.get('observaciones',''),
         user, estado_final, request.remote_addr))  # firmante = sesión autenticada (no spoofable por payload)
    # FIX 13-jun (audit recepción · M23): el kardex usa el estado CANÓNICO 'VIGENTE'
    # (no 'APROBADO'). 'APROBADO' es la etiqueta del review (queda en cc_reviews.estado_final);
    # escribirla en movimientos.estado_lote hacía que el cron de vencidos y los KPIs (filtros
    # de inclusión ='VIGENTE') SALTARAN el lote → consumo de vencido (M25) + stock fantasma.
    # Los otros 3 paths de liberación ya escriben VIGENTE.
    _estado_kardex = 'VIGENTE' if estado_final == 'APROBADO' else estado_final
    c.execute("UPDATE movimientos SET estado_lote=? WHERE id=?", (_estado_kardex, mov_id))
    # Ubicación final al LIBERAR (Sebastián 6-jun-2026): el flujo real es
    # "cuando Calidad libera, el lote pasa de la estantería CUARENTENA a su
    # ubicación final". Si al aprobar se indica estante/posición final, se mueve
    # el lote (TODAS sus filas de movimientos) → deja de estar en CUARENTENA física.
    ubic_final_msg = ''
    est_final = (d.get('estanteria_final') or '').strip()
    pos_final = (d.get('posicion_final') or '').strip()
    if estado_final == 'APROBADO' and (est_final or pos_final):
        _sets, _ps = [], []
        if est_final:
            _sets.append('estanteria=?'); _ps.append(est_final[:50])
        if pos_final:
            _sets.append('posicion=?'); _ps.append(pos_final[:50])
        c.execute(
            "UPDATE movimientos SET " + ', '.join(_sets) + " WHERE material_id=? AND lote=?",
            _ps + [mov[1], mov[2]])
        ubic_final_msg = ' · ubicación final: ' + (est_final or '—') + (('/' + pos_final) if pos_final else '')
    c.execute(
        "INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha) VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))",
        (user, 'CC_REVIEW_'+estado_final, 'movimientos', str(mov_id),
         'Lote '+d.get('lote','')+' AQL:'+resultado_aql+' Solub:'+solubilidad+' Firma:'+user+' e-sign #'+str(sig_id)+ubic_final_msg,
         request.remote_addr))
    if estado_final == 'RECHAZADO':
        # Fix #2 · 21-may-2026 · schema actualizado de solicitudes_compra.
        # Antes usaba columnas fantasma (material_codigo, cantidad, unidad...)
        # que ya no existen · INSERT fallaba silencioso · Catalina no se
        # enteraba del lote rechazado · zombie en bodega.
        try:
            from datetime import datetime as _dt
            anio_qc = _dt.now().strftime('%Y')
            cod_mp = d.get('codigo_mp', '')
            lote_rech = d.get('lote', '')
            nombre_mp = ''
            try:
                r_mp = c.execute(
                    "SELECT nombre_comercial FROM maestro_mps WHERE codigo_mp=?",
                    (cod_mp,),
                ).fetchone()
                if r_mp:
                    nombre_mp = r_mp[0] or ''
            except Exception:
                pass
            # numero único · DEV-YYYY-NNNN
            row_n = c.execute(
                "SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) "
                "FROM solicitudes_compra WHERE numero LIKE ?",
                (f'DEV-{anio_qc}-%',),
            ).fetchone()
            sol_num = f'DEV-{anio_qc}-{(row_n[0] or 0)+1:04d}'
            obs = f'LOTE RECHAZADO QC · devolución proveedor · lote: {lote_rech}'
            c.execute(
                """INSERT INTO solicitudes_compra
                   (numero, fecha, estado, solicitante, urgencia, observaciones,
                    area, empresa, categoria, tipo, valor)
                   VALUES (?, datetime('now','-5 hours'), 'Pendiente', ?,
                           'Alta', ?, 'Calidad', 'Espagiria', 'Materia Prima',
                           'Devolucion', 0)""",
                (sol_num, user or 'qc_user', obs),
            )
            try:
                c.execute(
                    """INSERT INTO solicitudes_compra_items
                       (numero, codigo_mp, nombre_mp, cantidad_g, valor_estimado)
                       VALUES (?, ?, ?, 0, 0)""",
                    (sol_num, cod_mp, nombre_mp or cod_mp),
                )
            except Exception:
                pass
            try:
                audit_log(c, usuario=user, accion='SOL_DEVOLUCION_QC',
                          tabla='solicitudes_compra', registro_id=sol_num,
                          despues={'lote': lote_rech, 'mp': cod_mp})
            except Exception:
                pass
        except Exception as _e:
            __import__('logging').getLogger('inventario').error(
                "Auto-creación de solicitud devolución falló: %s", _e
            )
    conn.commit()
    msgs = {'APROBADO': 'Lote APROBADO. Disponible para produccion.' + ubic_final_msg,
            'RECHAZADO': 'Lote RECHAZADO. Notificacion creada en Compras.',
            'CUARENTENA_EXTENDIDA': 'CUARENTENA EXTENDIDA. Maximo 5 dias para definicion.'}
    return jsonify({'message': msgs.get(estado_final,''), 'estado': estado_final})

@bp.route('/api/movimientos/<int:mov_id>/anular', methods=['POST'])
def anular_movimiento(mov_id):
    """Anula un movimiento generando un contra-movimiento. Requiere autenticacion."""
    user = session.get('compras_user', '')
    # SEC-FIX · 21-may-2026 · bloquear bypass user vacío
    # Antes: user='' pasaba el check `operador != user` con operador=''
    # Resultado: usuario sin sesión anulaba movimientos legacy
    if not user:
        return jsonify({'error': 'No autorizado'}), 401
    d = request.json or {}
    motivo = d.get('motivo', '').strip()
    if not motivo:
        return jsonify({'error': 'Motivo de anulacion requerido'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT m.*, mp.nombre FROM movimientos m
                 LEFT JOIN maestro_mps mp ON m.material_id = mp.codigo_mp
                 WHERE m.id=?""", (mov_id,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'Movimiento no encontrado'}), 404
    cols = [d[0] for d in c.description]
    mov = dict(zip(cols, row))
    # Verificar que no esté ya anulado
    if mov.get('observaciones','').startswith('[ANULADO]'):
        return jsonify({'error': 'Movimiento ya anulado'}), 400
    # Solo admin puede anular movimientos de otros usuarios.
    # El operario que registró el movimiento vive en la columna 'operador'
    # (movimientos no tiene 'responsable').
    if user not in ADMIN_USERS and mov.get('operador','') != user:
        return jsonify({'error': 'Solo puedes anular tus propios movimientos o ser administrador'}), 403
    # Generar contra-movimiento (invierte el tipo). Mismo lote y estado_lote
    # que el original para que el kardex (_get_mp_stock) lo concilie bien.
    tipo_inv = 'Salida' if mov['tipo'] == 'Entrada' else 'Entrada'
    obs_contra = f'[ANULACION] del movimiento #{mov_id} — {motivo} — por {user}'
    c.execute("""INSERT INTO movimientos
                 (material_id, material_nombre, tipo, cantidad, lote, estado_lote, operador, observaciones, fecha)
                 VALUES (?,?,?,?,?,?,?,?,datetime('now', '-5 hours'))""",
              (mov['material_id'], mov.get('material_nombre',''), tipo_inv,
               mov['cantidad'], mov.get('lote',''), mov.get('estado_lote',''),
               user, obs_contra))
    # Marcar original como anulado
    c.execute("UPDATE movimientos SET observaciones=? WHERE id=?",
              ('[ANULADO] ' + (mov.get('observaciones') or ''), mov_id))
    # Registrar en audit_log
    c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                 VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
              (user, 'ANULAR_MOVIMIENTO', 'movimientos', str(mov_id),
               f'Anulado mov #{mov_id} ({mov["tipo"]} {mov["cantidad"]}g de {mov["material_id"]}) — {motivo}',
               request.remote_addr))
    conn.commit()
    return jsonify({'ok': True, 'message': f'Movimiento #{mov_id} anulado. Contra-movimiento generado.',
                    'tipo_contramovimiento': tipo_inv})

@bp.route('/api/reset-movimientos', methods=['POST'])
def reset_mov():
    """OPERACIÓN PELIGROSA: borra TODOS los movimientos de inventario.

    Solo admins. Triple confirmación:
      1. Sesión debe ser ADMIN_USERS
      2. JSON.confirmacion == 'BORRAR_TODO_INVENTARIO_AHORA' (no solo 'BORRAR')
      3. JSON.fecha_actual == fecha de hoy (YYYY-MM-DD) — anti copy-paste accidental
    Antes de ejecutar: snapshot a backup_log + log de seguridad.
    """
    user = session.get('compras_user', '')
    if not user or user not in ADMIN_USERS:
        return jsonify({'error': 'No autorizado. Solo administradores.'}), 403

    d = request.json or {}
    if d.get('confirmacion', '') != 'BORRAR_TODO_INVENTARIO_AHORA':
        return jsonify({
            'error': 'confirmacion debe ser exactamente "BORRAR_TODO_INVENTARIO_AHORA"',
            'hint': 'Anti copy-paste accidental. Esta operación borra TODOS los movimientos.'
        }), 400

    from datetime import date as _date
    hoy = _date.today().isoformat()
    if d.get('fecha_actual', '') != hoy:
        return jsonify({
            'error': f'fecha_actual debe ser "{hoy}" (formato YYYY-MM-DD)',
            'hint': 'Doble confirmación: la fecha de HOY debe coincidir.'
        }), 400

    # Forzar backup antes de borrar — recuperabilidad obligatoria
    try:
        from backup import do_backup
        bk_result = do_backup(triggered_by=f"pre-reset:{user}")
        if not bk_result.get('ok') and not bk_result.get('skipped'):
            return jsonify({
                'error': 'No se pudo crear backup pre-reset. Operación abortada.',
                'detail': bk_result.get('error', '')[:200]
            }), 500
        backup_filename = bk_result.get('filename', '(skipped — otro en curso)')
    except Exception as e:
        return jsonify({
            'error': 'Backup pre-reset falló. Operación abortada.',
            'detail': str(e)[:200]
        }), 500

    # Log de seguridad ANTES (en caso de crash queda traza del intento)
    from auth import _log_sec, _client_ip
    _log_sec("inventario_reset_INICIADO", user, _client_ip(),
             f"backup={backup_filename}")

    conn = get_db(); c = conn.cursor()
    rows_deleted = c.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
    c.execute("DELETE FROM movimientos")
    c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                 VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
              (user, 'RESET_MOVIMIENTOS', 'movimientos', 'ALL',
               f'Borrados {rows_deleted} movimientos. Backup pre-reset: {backup_filename}',
               _client_ip()))
    conn.commit()

    _log_sec("inventario_reset_COMPLETADO", user, _client_ip(),
             f"deleted={rows_deleted} backup={backup_filename}")

    return jsonify({
        'ok': True,
        'message': f'{rows_deleted} movimientos borrados.',
        'backup_pre_reset': backup_filename,
        'restore_hint': 'Si fue un error: descarga el backup desde /admin → Backups y restaura.'
    })

def _fefo_lote_rotulo(conn, material_id, material_nombre=''):
    """Lote FEFO que producción usaría para este MP: código RESUELTO + lote con
    stock neto>0 en estado producible, el de vencimiento más próximo. '' si no hay.
    Usado por el rótulo de pesaje y la hoja de pesaje del detalle de orden · que
    el lote mostrado sea SIEMPRE el que se va a descontar."""
    try:
        from blueprints.programacion import (_resolver_material_bodega as _res,
                                             _ESTADOS_LOTE_NO_PRODUCIBLES as _NP)
    except Exception:
        return ''
    cod = material_id
    try:
        cod = _res(conn, material_id, material_nombre) or material_id
    except Exception:
        cod = material_id
    _ph = ','.join(['?'] * len(_NP))
    try:
        row = conn.execute(
            f"""SELECT lote FROM movimientos
                WHERE material_id=? AND COALESCE(lote,'') NOT IN ('','S/L')
                  AND UPPER(COALESCE(estado_lote,'')) NOT IN ({_ph})
                GROUP BY lote
                HAVING SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) > 0
                  AND (MAX(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA') THEN fecha_vencimiento END) IS NULL
                       OR TRIM(CAST(MAX(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA') THEN fecha_vencimiento END) AS TEXT))=''
                       OR date(MAX(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA') THEN fecha_vencimiento END)) >= date('now', '-5 hours'))
                ORDER BY COALESCE(NULLIF(CAST(MAX(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA') THEN fecha_vencimiento END) AS TEXT),''),'9999-12-31') ASC
                LIMIT 1""",
            (cod,) + tuple(_NP)).fetchone()
        return row[0] if row else ''
    except Exception:
        return ''


@bp.route('/rotulos/<producto_nombre>/<cantidad_str>')
def generar_rotulos(producto_nombre, cantidad_str):
    if 'compras_user' not in session:
        return redirect('/login')
    try: cantidad_kg = float(cantidad_str)
    except: return "<h2>Cantidad invalida</h2>", 400
    from datetime import date; import urllib.parse
    hoy = date.today().strftime('%d-%b-%Y').upper()
    prod = urllib.parse.unquote(producto_nombre); op_num = "OP-"+date.today().strftime('%Y%m%d'); cant_g = cantidad_kg*1000
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT material_id,material_nombre,porcentaje FROM formula_items WHERE producto_nombre=?", (prod,))
    items = c.fetchall()
    # Audit 5-jun-2026 · el LOTE del rótulo debe ser el MISMO que producción
    # descuenta (FEFO real): código RESUELTO (no el crudo de fórmula), con stock
    # NETO > 0 y en estado producible (excluye CUARENTENA/AGOTADO/VENCIDO/...).
    # Antes usaba el código crudo + cualquier Entrada → podía mostrar S/L (cuando
    # el stock está bajo un código duplicado) o un lote bloqueado/agotado.
    try:
        from blueprints.programacion import (_resolver_material_bodega as _res_rot,
                                             _ESTADOS_LOTE_NO_PRODUCIBLES as _NP_ROT)
    except Exception:
        _res_rot, _NP_ROT = None, None
    lotes = {}; incis = {}; cods = {}
    for r in items:
        mid = r[0]; mnm = r[1] if len(r) > 1 else ''
        c.execute("SELECT nombre_inci FROM maestro_mps WHERE codigo_mp=?", (mid,)); ir=c.fetchone(); incis[mid]=ir[0] if ir and ir[0] else ''
        cod = mid
        if _res_rot:
            try: cod = _res_rot(c, mid, mnm) or mid
            except Exception: cod = mid
        cods[mid] = cod
        row = None
        if _NP_ROT:
            _ph = ','.join(['?'] * len(_NP_ROT))
            try:
                row = c.execute(
                    f"""SELECT lote, MAX(estanteria), MAX(posicion),
                              MAX(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA') THEN fecha_vencimiento END) AS fv,
                              SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) AS stk
                       FROM movimientos
                       WHERE material_id=? AND COALESCE(lote,'') NOT IN ('','S/L')
                         AND UPPER(COALESCE(estado_lote,'')) NOT IN ({_ph})
                       GROUP BY lote HAVING stk > 0
                         AND (fv IS NULL OR TRIM(CAST(fv AS TEXT))='' OR date(fv) >= date('now', '-5 hours'))
                       ORDER BY COALESCE(NULLIF(CAST(MAX(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA') THEN fecha_vencimiento END) AS TEXT),''),'9999-12-31') ASC
                       LIMIT 1""",
                    (cod,) + tuple(_NP_ROT)).fetchone()
            except Exception:
                row = None
        if row is None:
            # Fallback legacy (sin resolver/estados) · PG-compat CAST a TEXT.
            c.execute("SELECT lote,estanteria,posicion,fecha_vencimiento FROM movimientos WHERE material_id=? AND tipo='Entrada' AND lote IS NOT NULL AND lote!='' AND lote!='S/L' ORDER BY COALESCE(NULLIF(CAST(fecha_vencimiento AS TEXT),''),'9999-12-31') ASC LIMIT 1", (cod,))
            row=c.fetchone()
        lotes[mid]={'lote':row[0] if row else 'S/L','est':row[1] if row else '','pos':row[2] if row else '','vence':str(row[3])[:10] if row and row[3] else ''}
    if not items: return '<h2>Formula no encontrada: '+prod+'</h2>', 404
    rhtml=''; barcodes=''
    for i,r in enumerate(items):
        mid,mnm,pct=r; peso=round((pct/100)*cant_g,2); info=lotes.get(mid,{}); lote_mp=info.get('lote','S/L')
        cod_real=cods.get(mid,mid)
        ubicacion=('Est. '+str(info.get('est',''))+str(info.get('pos',''))).strip(); vence=info.get('vence',''); inci=incis.get(mid,'')
        bv=cod_real+'|'+lote_mp; barcodes+=f'try{{JsBarcode("#bc{i}","{bv}",{{format:"CODE128",width:1.2,height:35,displayValue:false,margin:0}})}}catch(e){{}};'
        rhtml+='<div class="r"><div class="rh"><span class="rt">RÓTULO PARA DISPENSAR MATERIA PRIMA</span><span class="rc">PRD-PRO-001-F08 | v1<br>04-Mar-2025 / 03-Mar-2028</span></div>'
        rhtml+='<table><tr><td class="l">OP:</td><td class="v">'+op_num+'</td><td class="l">Fecha:</td><td class="v">'+hoy+'</td></tr>'
        rhtml+='<tr><td class="l">Producto:</td><td class="v big" colspan="3"><b>'+prod+'</b> &mdash; '+str(cantidad_kg)+' kg</td></tr>'
        rhtml+='<tr><td class="l">Nombre MP:</td><td class="v bold" colspan="3"><b>'+mnm+'</b> <span style="color:#888;font-size:0.8em;">('+mid+')</span></td></tr>'
        if inci: rhtml+='<tr><td class="l">Nombre INCI:</td><td class="v" colspan="3" style="font-size:0.82em;color:#444;">'+inci+'</td></tr>'
        rhtml+='<tr><td class="l">Lote MP:</td><td class="v bold">'+lote_mp+'</td><td class="l">Ubicacion:</td><td class="v">'+ubicacion+'</td></tr>'
        rhtml+='<tr><td class="l">Vencimiento:</td><td class="v" style="color:#c0392b;">'+vence+'</td><td class="l">% formula:</td><td class="v">'+str(pct)+'%</td></tr>'
        rhtml+='<tr><td class="l">Peso teorico:</td><td class="v peso">'+f"{peso:,.2f} g"+'</td><td class="l">Lote Prod.:</td><td class="blank"></td></tr>'
        rhtml+='<tr><td class="l">Tara:</td><td class="blank"></td><td class="l">Peso Neto:</td><td class="blank"></td></tr>'
        rhtml+='<tr><td class="l">Pesado por:</td><td class="blank firma"></td><td class="l">Verificado:</td><td class="blank firma"></td></tr>'
        rhtml+='</table>'
        rhtml+='<div style="text-align:center;padding:4px;"><svg id="bc'+str(i)+'"></svg></div>'
        rhtml+='<div class="rf">'+cod_real+'|'+lote_mp+' | #'+str(i+1)+' de '+str(len(items))+'</div></div>'
    css=('<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script><title>Rotulos</title>'
         '<script src="https://cdnjs.cloudflare.com/ajax/libs/jsbarcode/3.11.5/JsBarcode.all.min.js"></script>'
         '<style>*{margin:0;padding:0;box-sizing:border-box;}body{font-family:Arial,sans-serif;font-size:9pt;background:#eee;}'
         '.ph{background:#4c1d95;color:white;padding:10px 16px;display:flex;justify-content:space-between;align-items:center;}'
         '.pbtn{background:#6d28d9;color:white;border:none;padding:8px 20px;border-radius:8px;cursor:pointer;font-weight:600;}'
         '.wrap{display:flex;flex-wrap:wrap;gap:5px;padding:8px;}'
         '.r{background:white;border:2px solid #4c1d95;border-radius:3px;width:370px;page-break-inside:avoid;}'
         '.rh{background:#4c1d95;color:white;padding:5px 8px;display:flex;justify-content:space-between;align-items:center;}'
         '.rt{font-weight:bold;font-size:8pt;}.rc{font-size:6.5pt;text-align:right;line-height:1.4;}'
         'table{width:100%;border-collapse:collapse;}td{border:1px solid #bbb;padding:3px 5px;vertical-align:middle;}'
         '.l{background:#ecf0f1;font-weight:bold;font-size:7.5pt;color:#4c1d95;white-space:nowrap;width:27%;}'
         '.v{font-size:8.5pt;width:23%;}.bold{font-size:9pt;}.big{font-size:9pt;}'
         '.peso{background:#fff3cd;color:#c0392b;font-size:12pt;font-weight:bold;}'
         '.blank{height:20px;width:23%;}.firma{height:26px;}.rf{background:#ecf0f1;padding:2px 6px;font-size:6.5pt;color:#888;text-align:right;}'
         '@media print{body{background:white;}.ph{display:none;}.wrap{padding:0;gap:3px;}.r{width:48%;}@page{size:letter landscape;margin:7mm;}}'
         '</style></head><body>')
    return (css+'<div class="ph"><div><h2>Rotulos &mdash; '+prod+' &mdash; '+str(cantidad_kg)+' kg</h2>'
            '<div style="font-size:8pt;opacity:0.8;">'+op_num+' | '+str(len(items))+' MPs | '+hoy+'</div></div>'
            '<button class="pbtn" onclick="window.print()">Imprimir todos</button></div>'
            '<div class="wrap">'+rhtml+'</div>'
            '<script>window.onload=function(){'+barcodes+'};</script>'
            '</body></html>')

@bp.route('/rotulo-recepcion/<codigo>/<lote>/<cantidad_str>')
def rotulo_recepcion(codigo, lote, cantidad_str):
    if 'compras_user' not in session:
        return redirect('/login')
    try: cantidad = float(cantidad_str)
    except: return "<h2>Cantidad invalida</h2>", 400
    from datetime import date; import urllib.parse
    hoy = date.today().strftime('%d-%b-%Y').upper(); lote=urllib.parse.unquote(lote)
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT nombre_inci,nombre_comercial,tipo,proveedor FROM maestro_mps WHERE codigo_mp=?", (codigo,)); mp=c.fetchone()
    c.execute("SELECT fecha_vencimiento,estanteria,posicion,proveedor FROM movimientos WHERE material_id=? AND lote=? ORDER BY fecha DESC LIMIT 1", (codigo,lote)); mov=c.fetchone()
    ni=mp[0] if mp else ''; nc=mp[1] if mp else codigo; tp=mp[2] if mp else ''
    pv=(mp[3] if mp and mp[3] else '') or (mov[3] if mov and len(mov)>3 and mov[3] else '')
    fv=str(mov[0])[:10] if mov and mov[0] else ''; ub=((mov[1] or '')+(mov[2] or '')) if mov else ''
    nr="REC-"+date.today().strftime('%Y%m%d')+"-"+codigo[-3:]; bv=codigo+'|'+lote
    h=('<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
       '<title>Rotulo Recepcion MP</title>'
       '<script src="https://cdnjs.cloudflare.com/ajax/libs/jsbarcode/3.11.5/JsBarcode.all.min.js"></script>'
       '<style>'
       '*{margin:0;padding:0;box-sizing:border-box;}'
       'body{font-family:Arial,sans-serif;font-size:10pt;background:#eee;padding:20px;}'
       '.ph{background:#4c1d95;color:white;padding:10px 16px;display:flex;justify-content:space-between;margin-bottom:10px;}'
       '.pb{background:#6d28d9;color:white;border:none;padding:8px 20px;border-radius:8px;cursor:pointer;font-weight:600;}'
       '.r{background:white;border:3px solid #4c1d95;border-radius:5px;max-width:520px;margin:auto;}'
       '.rh{background:#4c1d95;color:white;padding:8px 12px;text-align:center;}'
       '.lote{background:#fff3cd;border:2px solid #f39c12;padding:10px;text-align:center;margin:10px;}'
       '.lnum{font-size:20pt;font-weight:bold;color:#c0392b;letter-spacing:2px;}'
       'table{width:100%;border-collapse:collapse;}td{border:1px solid #ccc;padding:6px 8px;}'
       '.l{background:#ecf0f1;font-weight:bold;font-size:8.5pt;width:35%;}'
       '.termica{display:none;}'
       '@media print{'
         '.ph{display:none!important;}.r{display:none!important;}'
         '.termica{display:block!important;}'
         'body{background:white;padding:0;margin:0;}'
         '@page{size:50mm auto;margin:1mm;}'
       '}'
       '</style></head><body>'
    )
    h+=('<div class="ph"><b>Rotulo de Recepcion Materia Prima</b>'
        '<button class="pb" onclick="window.print()">&#128438; Imprimir etiqueta termica</button></div>'
    )
    h+=('<div class="r"><div class="rh">'
        '<span style="font-weight:bold;font-size:11pt;display:block;margin-bottom:2px;">ROTULO DE INGRESO DE MATERIA PRIMA</span>'
        '<span style="font-size:7.5pt;opacity:0.85;">Espagiria Laboratorios &nbsp;|&nbsp; COC-PRO-002-F07 &nbsp;|&nbsp; '+hoy+'</span>'
        '</div>'
        '<div class="lote"><div style="font-size:9pt;color:#888;margin-bottom:4px;">NUMERO DE LOTE</div>'
        '<div class="lnum">'+lote+'</div>'
        '<svg id="bc" style="margin-top:6px;"></svg>'
        '<div style="font-size:7pt;color:#888;margin-top:2px;">'+bv+'</div></div>'
        '<table>'
        '<tr><td class="l">Codigo MP:</td><td style="font-weight:700;">'+codigo+'</td></tr>'
        '<tr><td class="l">Nombre INCI:</td><td style="font-size:0.9em;color:#1a5276;">'+ni+'</td></tr>'
        '<tr><td class="l">Nombre Comercial:</td><td style="font-weight:700;">'+nc+'</td></tr>'
        '<tr><td class="l">Tipo / Funcion:</td><td>'+tp+'</td></tr>'
        '<tr><td class="l">Proveedor:</td><td style="font-weight:700;">'+pv+'</td></tr>'
        '<tr><td class="l">Cantidad recibida:</td><td style="color:#27ae60;font-weight:700;">'+f"{cantidad:,.0f} g"+'</td></tr>'
        '<tr><td class="l">Fecha de recepcion:</td><td style="font-weight:700;">'+hoy+'</td></tr>'
        '<tr><td class="l">Fecha de vencimiento:</td><td style="color:#c0392b;font-weight:700;">'+fv+'</td></tr>'
        '<tr><td class="l">Fecha de analisis:</td><td style="height:28px;background:#fffde7;"></td></tr>'
        '<tr style="background:#e8f5e9;">'
        '<td class="l" style="color:#1b5e20;font-weight:800;font-size:10pt;vertical-align:middle;">Estado de calidad:</td>'
        '<td style="height:70px;vertical-align:top;padding:8px;">'
        '<div style="display:flex;gap:20px;margin-bottom:6px;">'
        '<span style="font-size:11pt;font-weight:700;">&#9744; Aprobado</span>'
        '<span style="font-size:11pt;font-weight:700;">&#9744; Cuarentena</span>'
        '<span style="font-size:11pt;font-weight:700;">&#9744; Rechazado</span>'
        '</div>'
        '<div style="background:#fffde7;border:1px dashed #f39c12;height:36px;border-radius:3px;display:flex;align-items:center;justify-content:center;">'
        '<span style="font-size:7.5pt;color:#aaa;">[ espacio para sticker de calidad ]</span>'
        '</div></td></tr>'
        '<tr><td class="l">Ubicacion:</td><td>Est. '+ub+'</td></tr>'
        '<tr><td class="l">N de Recepcion:</td><td>'+nr+'</td></tr>'
        '<tr><td class="l">Recibido por:</td><td style="height:30px;"></td></tr>'
        '<tr><td class="l">Analizado / Aprobado por:</td><td style="height:30px;"></td></tr>'
        '</table>'
        '<div style="background:#ecf0f1;padding:4px 10px;font-size:7.5pt;color:#888;text-align:center;">'
        'COC-PRO-002-F07 &nbsp;|&nbsp; '+hoy+'</div></div>'
    )
    h+=('<div class="termica" style="width:50mm;font-family:Arial,sans-serif;font-size:6.5pt;border:2px solid #000;word-break:break-word;">'
        '<div style="background:#000;color:#fff;text-align:center;padding:2px;font-size:6pt;font-weight:bold;">ESPAGIRIA LAB &nbsp;|&nbsp; COC-PRO-002-F07</div>'
        '<div style="text-align:center;padding:3px 2px 1px;">'
        '<svg id="bc2" style="width:46mm;height:18mm;"></svg>'
        '<div style="font-size:8pt;font-weight:bold;color:#c0392b;letter-spacing:1px;margin-top:1px;">'+lote+'</div>'
        '</div>'
        '<table style="width:100%;border-collapse:collapse;font-size:6.5pt;">'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;width:40%;">Codigo:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-weight:bold;">'+codigo+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">INCI:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-style:italic;">'+ni+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Tipo/Funcion:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;">'+tp+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Proveedor:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-weight:bold;">'+pv+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Nombre:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-weight:bold;">'+nc+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Cantidad:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-weight:bold;color:#27ae60;">'+f"{cantidad:,.0f} g"+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Recep:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;">'+hoy+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Vence:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-weight:bold;color:#c0392b;">'+fv+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Analisis:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;background:#fffde7;height:10px;"></td></tr>'
        '<tr><td colspan="2" style="border:1px solid #999;padding:1px 2px;background:#c8f7c5;font-weight:bold;">Estado de Calidad:</td></tr>'
        '<tr><td colspan="2" style="border:1px solid #999;padding:2px 3px;background:#c8f7c5;">'
        '<span style="margin-right:6px;">&#9744; Aprobado</span>'
        '<span style="margin-right:6px;">&#9744; Cuarentena</span>'
        '<span>&#9744; Rechazado</span>'
        '<div style="margin-top:2px;border:1px dashed #f39c12;background:#fffde7;height:16px;'
        'display:flex;align-items:center;justify-content:center;">'
        '<span style="font-size:5.5pt;color:#aaa;">[sticker QC]</span></div></td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Ubicacion:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;font-weight:bold;">Est. '+ub+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">N Recepcion:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;">'+nr+'</td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Recibido por:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;height:14px;"></td></tr>'
        '<tr><td style="border:1px solid #999;padding:1px 2px;background:#eee;font-weight:bold;">Anal./Aprobado:</td>'
        '<td style="border:1px solid #999;padding:1px 2px;height:14px;"></td></tr>'
        '</table>'
        '<div style="text-align:center;font-size:5.5pt;color:#666;padding:1px;">'+nr+'</div>'
        '</div>'
    )
    h+=('<script>window.onload=function(){'
        'try{JsBarcode("#bc","'+bv+'",{format:"CODE128",width:1.5,height:45,displayValue:false,margin:0});}catch(e){}'
        'try{JsBarcode("#bc2","'+bv+'",{format:"CODE128",width:1,height:35,displayValue:false,margin:0});}catch(e){}'
        '}</script>'
        '</body></html>')
    return h

@bp.route('/rotulo-recepcion-mee/<codigo>/<cantidad_str>')
def rotulo_recepcion_mee(codigo, cantidad_str):
    if 'compras_user' not in session:
        return redirect('/login')
    try: cantidad = int(float(cantidad_str))
    except: return "<h2>Cantidad invalida</h2>", 400
    from datetime import date; import urllib.parse
    hoy = date.today().strftime('%d-%b-%Y').upper()
    codigo = urllib.parse.unquote(codigo)
    lote_ref = request.args.get('lote','')
    try:
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT descripcion, categoria, proveedor, unidad FROM maestro_mee WHERE codigo=?", (codigo,))
        mee = c.fetchone()
        c.execute("SELECT lote_ref, responsable, fecha, observaciones FROM movimientos_mee WHERE mee_codigo=? AND tipo='Entrada' AND anulado=0 ORDER BY id DESC LIMIT 1", (codigo,))
        mov = c.fetchone()
    except Exception as e:
        return f"<h2>Error DB: {e}</h2>", 500
    desc  = mee[0] if mee else codigo
    cat   = mee[1] if mee else ''
    prov  = mee[2] if mee else ''
    unid  = mee[3] if mee and len(mee)>3 else 'und'
    lote  = lote_ref or (mov[0] if mov else '')
    oper  = mov[1] if mov else ''
    obs   = mov[3] if mov and len(mov)>3 else ''
    nr    = "REC-MEE-" + date.today().strftime('%Y%m%d') + "-" + codigo[-4:]
    # M.ENV: envases primarios / M.EMP: empaque secundario
    env_cats = {'Envase','Frasco','Tapa','Gotero','Contorno'}
    is_env = cat in env_cats
    chk_mp  = '&#9744;'; chk_env = '&#9745;' if is_env else '&#9744;'; chk_emp = '&#9745;' if not is_env else '&#9744;'
    cant_str = f"{cantidad:,} {unid}"

    css = ('<style>'
           '*{margin:0;padding:0;box-sizing:border-box;}'
           'body{font-family:Arial,sans-serif;font-size:9pt;background:#ddd;padding:16px;}'
           '.ph{background:#1a3a5c;color:white;padding:8px 16px;display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;}'
           '.pb{background:#2980b9;color:white;border:none;padding:6px 18px;border-radius:4px;cursor:pointer;font-weight:bold;}'
           '.grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;max-width:920px;margin:auto;}'
           '.rot{background:white;border:2.5px solid #1a3a5c;border-radius:4px;}'
           '.rh{background:#1a3a5c;color:white;padding:5px 10px;display:grid;grid-template-columns:1fr auto;gap:2px;font-size:8pt;}'
           '.rh-title{font-weight:700;font-size:9.5pt;}'
           '.rh-meta{text-align:right;font-size:7pt;opacity:0.85;line-height:1.5;}'
           'table{width:100%;border-collapse:collapse;}'
           'td{border:1px solid #bbb;padding:3px 6px;font-size:8.5pt;}'
           '.lbl{background:#ecf0f1;font-weight:700;font-size:7.5pt;color:#333;width:42%;}'
           '.val{font-weight:600;}'
           '.tipo{background:#dbe9f5;}'
           '.calidad td{background:#e8f5e9;}'
           '.firma td{height:30px;}'
           '.bc{text-align:center;padding:5px;background:#f8f9fa;border-bottom:1px solid #bbb;}'
           '.footer{background:#dde8f0;padding:3px 8px;font-size:7pt;color:#555;text-align:center;}'
           '@media print{'
           '@page{size:A4;margin:8mm;}'
           '.ph{display:none;}'
           'body{background:white;padding:0;}'
           '.grid{display:grid;grid-template-columns:1fr 1fr;gap:4mm;}'
           '.rot{border:1.5px solid #000;page-break-inside:avoid;}'
           '}'
           '</style>')

    def make_label(bid):
        lbl  = '<div class="rot">'
        lbl += ('<div class="rh">'
                '<div><div class="rh-title">IDENTIFICACI&Oacute;N DE INSUMOS</div>'
                '<div style="font-size:7pt;opacity:0.8;">Espagiria Laboratorios</div></div>'
                '<div class="rh-meta">'
                'C&oacute;digo: <b>COC-PRO-002-F04</b><br>'
                'Versi&oacute;n: 2 &nbsp;|&nbsp; P&aacute;g: 1 de 1<br>'
                'Vigencia: 13-Jun-2025 / 12-Jun-2028'
                '</div></div>')
        lbl += (f'<div class="bc">'
                f'<div style="font-size:7pt;color:#666;margin-bottom:2px;">C&Oacute;DIGO &mdash; BARRAS</div>'
                f'<svg id="{bid}" style="max-width:100%;"></svg>'
                f'<div style="font-family:monospace;font-weight:700;font-size:9pt;letter-spacing:2px;color:#1a3a5c;">{codigo}</div>'
                f'</div>')
        lbl += '<table>'
        lbl += f'<tr><td class="lbl">NOMBRE COMERCIAL DEL INSUMO</td><td class="val" colspan="3">{desc}</td></tr>'
        lbl += f'<tr><td class="lbl">NOMBRE INCI DEL INSUMO</td><td colspan="3">&mdash;</td></tr>'
        lbl += f'<tr><td class="lbl">MARCA O FORMA QU&Iacute;MICA</td><td colspan="3">{prov}</td></tr>'
        lbl += ('<tr class="tipo"><td class="lbl tipo">TIPO DE INSUMO</td>'
                f'<td style="text-align:center;width:18%;">{chk_mp} MP</td>'
                f'<td style="text-align:center;width:18%;">{chk_env} M.ENV</td>'
                f'<td style="text-align:center;width:18%;">{chk_emp} M.EMP</td></tr>')
        lbl += f'<tr><td class="lbl">C&Oacute;DIGO INTERNO</td><td class="val">{codigo}</td><td class="lbl">LOTE</td><td class="val">{lote}</td></tr>'
        lbl += f'<tr><td class="lbl">CANTIDAD</td><td class="val">{cant_str}</td><td class="lbl">PROVEEDOR</td><td class="val">{prov}</td></tr>'
        lbl += f'<tr><td class="lbl">FECHA DE RECEPCI&Oacute;N</td><td class="val">{hoy}</td><td class="lbl">FECHA DE AN&Aacute;LISIS</td><td style="height:26px;background:#fffde7;"></td></tr>'
        lbl += f'<tr><td class="lbl">OBSERVACIONES</td><td colspan="3" style="height:24px;">{obs}</td></tr>'
        lbl += f'<tr><td class="lbl">FECHA DE VENCIMIENTO</td><td colspan="3">N/A &mdash; Material de envase/empaque</td></tr>'
        lbl += ('<tr class="calidad"><td class="lbl calidad" style="color:#1b5e20;">ESTADO</td>'
                '<td colspan="3" style="height:26px;">'
                '<span style="margin-right:12px;">&#9744; Aprobado</span>'
                '<span style="margin-right:12px;">&#9744; En cuarentena</span>'
                '<span>&#9744; Rechazado</span></td></tr>')
        lbl += '<tr class="firma"><td class="lbl">FECHA Y FIRMA REALIZADO POR</td><td colspan="3"></td></tr>'
        lbl += '<tr class="firma"><td class="lbl">FECHA Y FIRMA APROBADO POR</td><td colspan="3"></td></tr>'
        lbl += '</table>'
        lbl += f'<div class="footer">COC-PRO-002-F04 &nbsp;|&nbsp; {cat} &nbsp;|&nbsp; {hoy} &nbsp;|&nbsp; N&deg; Rec: {nr}</div>'
        lbl += '</div>'
        return lbl

    h  = '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
    h += '<script src="https://cdnjs.cloudflare.com/ajax/libs/jsbarcode/3.11.5/JsBarcode.all.min.js"></script>'
    h += css + '</head><body>'
    h += ('<div class="ph">'
          '<b>R&oacute;tulo Identificaci&oacute;n Insumos MEE &mdash; Espagiria '
          '<span style="font-size:0.8em;opacity:0.8;">(4 etiquetas por hoja A4)</span></b>'
          '<button class="pb" onclick="window.print()">&#128203; Imprimir 4</button></div>')
    h += '<div class="grid">'
    bids = ['bc0','bc1','bc2','bc3']
    for bid in bids:
        h += make_label(bid)
    h += '</div>'
    js_parts = [f'try{{JsBarcode("#{b}","{codigo}",{{format:"CODE128",width:1.6,height:40,displayValue:false,margin:1}});}}catch(e){{}}' for b in bids]
    h += '<script>window.onload=function(){' + ''.join(js_parts) + '}</script>'
    h += '</body></html>'
    return h
    return h

@bp.route('/api/ordenes-compra/pendientes-recepcion')
def ocs_pendientes_recepcion():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT oc.numero_oc, oc.proveedor, oc.fecha, oc.valor_total,
                        oci.codigo_mp, oci.nombre_mp, oci.cantidad_g, oci.precio_unitario
                 FROM ordenes_compra oc
                 JOIN ordenes_compra_items oci ON oc.numero_oc = oci.numero_oc
                 WHERE oc.estado IN ('Aprobada','Enviada','Parcial')
                 ORDER BY oc.fecha DESC""")
    rows = c.fetchall()
    ocs = {}
    for r in rows:
        num = r[0]
        if num not in ocs:
            ocs[num] = {'numero_oc': num, 'proveedor': r[1], 'fecha': r[2],
                        'valor_total': r[3], 'items': []}
        ocs[num]['items'].append({'codigo_mp': r[4], 'nombre_mp': r[5],
                                   'cantidad_g': r[6], 'precio_unitario': r[7]})
    return jsonify(list(ocs.values()))

@bp.route('/api/trazabilidad/lote/<path:lote>')
def trazabilidad_lote_path(lote):
    import urllib.parse; lote = urllib.parse.unquote(lote)
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT id, material_id, material_nombre, cantidad, tipo, fecha,
                        observaciones, proveedor, precio_kg, numero_factura, numero_oc
                 FROM movimientos WHERE lote=? ORDER BY fecha""", (lote,))
    cols = [d[0] for d in c.description]
    movs = [dict(zip(cols, r)) for r in c.fetchall()]
    c.execute("""SELECT id, producto, cantidad, fecha, observaciones, operador
                 FROM producciones WHERE observaciones LIKE ? ORDER BY fecha""", (f'%{lote}%',))
    cols2 = [d[0] for d in c.description]
    prods = [dict(zip(cols2, r)) for r in c.fetchall()]
    c.execute("""SELECT d.numero, cl.nombre, d.fecha, d.estado
                 FROM despachos d LEFT JOIN clientes cl ON d.cliente_id=cl.id
                 WHERE d.observaciones LIKE ?""", (f'%{lote}%',))
    cols3 = [d[0] for d in c.description]
    desps = [dict(zip(cols3, r)) for r in c.fetchall()]
    return jsonify({'lote': lote, 'movimientos': movs, 'producciones': prods, 'despachos': desps})

@bp.route('/api/mp/<codigo>/historial-precios')
def historial_precios_mp(codigo):
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT fecha, proveedor, precio_kg, valor_total, numero_factura, numero_oc
                 FROM movimientos WHERE material_id=? AND tipo='Entrada' AND precio_kg>0
                 ORDER BY fecha DESC LIMIT 24""", (codigo,))
    hist = [{'fecha':r[0],'proveedor':r[1],'precio_kg':r[2],'valor_total':r[3],'factura':r[4],'oc':r[5]} for r in c.fetchall()]
    c.execute("SELECT precio_referencia, proveedor FROM maestro_mps WHERE codigo_mp=?", (codigo,))
    mp = c.fetchone()
    return jsonify({'codigo': codigo, 'precio_referencia': mp[0] if mp else 0,
                    'proveedor_habitual': mp[1] if mp else '', 'historial': hist})

@bp.route('/api/mp/<codigo>/consumo-historico')
def consumo_historico_mp(codigo):
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT substr(fecha,1,7) as mes,
                        SUM(CASE WHEN tipo='Salida' THEN cantidad ELSE 0 END) as consumo_g,
                        COUNT(CASE WHEN tipo='Salida' THEN 1 END) as n_salidas
                 FROM movimientos WHERE material_id=?
                 GROUP BY substr(fecha,1,7) ORDER BY mes DESC LIMIT 12""", (codigo,))
    meses = [{'mes':r[0],'consumo_g':r[1],'n_salidas':r[2]} for r in c.fetchall()]
    consumos = [m['consumo_g'] for m in meses if m['consumo_g'] and m['consumo_g'] > 0]
    promedio = sum(consumos)/len(consumos) if consumos else 0
    c.execute("SELECT lead_time_dias, stock_minimo FROM maestro_mps WHERE codigo_mp=?", (codigo,))
    mp = c.fetchone()
    lead = (mp[0] if mp and mp[0] else 7)
    stock_min = (mp[1] if mp and mp[1] else 0)
    punto_reorden = (promedio/30) * lead + stock_min
    return jsonify({'codigo': codigo, 'meses': meses,
                    'promedio_mes_g': round(promedio, 0),
                    'consumo_diario_g': round(promedio/30, 1),
                    'lead_time_dias': lead,
                    'punto_reorden_g': round(punto_reorden, 0)})

@bp.route('/api/conteos', methods=['GET','POST'])
def conteos():
    # P0 audit 26-may-2026 · sin gate · creación/listado de conteos abierto.
    _u_sess, _err_s, _code_s = _require_session()
    if _err_s:
        return _err_s, _code_s
    if request.method == 'POST':
        _u_w, _err_w, _code_w = _require_planta_write()
        if _err_w:
            return _err_w, _code_w
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.json or {}
        num = 'CNT-' + datetime.now().strftime('%Y%m%d-%H%M')
        c.execute("""INSERT INTO conteos_fisicos (numero,fecha_inicio,estado,responsable,observaciones)
                     VALUES (?,?,'Abierto',?,?)""",
                  (num, datetime.now().isoformat(), d.get('responsable',''), d.get('observaciones','')))
        cid = c.lastrowid
        c.execute("""SELECT mp.codigo_mp, mp.nombre_comercial,
                            COALESCE(SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad
                                         WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END),0)
                     FROM maestro_mps mp
                     LEFT JOIN movimientos m ON mp.codigo_mp=m.material_id
                     WHERE mp.activo=1 GROUP BY mp.codigo_mp""")
        mps = c.fetchall()
        for mp in mps:
            c.execute("INSERT INTO conteo_items (conteo_id,codigo_mp,nombre_mp,stock_sistema) VALUES (?,?,?,?)",
                      (cid, mp[0], mp[1], max(0, mp[2])))
        c.execute("UPDATE conteos_fisicos SET total_items=? WHERE id=?", (len(mps), cid))
        conn.commit()
        return jsonify({'numero': num, 'id': cid, 'total_items': len(mps)}), 201
    c.execute("SELECT id,numero,fecha_inicio,estado,responsable,total_items,items_diferencia FROM conteos_fisicos ORDER BY fecha_inicio DESC LIMIT 20")
    rows = [{'id':r[0],'numero':r[1],'fecha':r[2],'estado':r[3],'responsable':r[4],'total':r[5],'diffs':r[6]} for r in c.fetchall()]
    return jsonify(rows)

@bp.route('/api/conteos/<int:cid>', methods=['GET','PATCH'])
def conteo_detalle(cid):
    # P0 audit 26-may-2026 · sin gate · PATCH aplicar_ajustes hacía INSERT a movimientos.
    _u_sess, _err_s, _code_s = _require_session()
    if _err_s:
        return _err_s, _code_s
    if request.method == 'PATCH':
        _u_w, _err_w, _code_w = _require_planta_write()
        if _err_w:
            return _err_w, _code_w
    conn = get_db(); c = conn.cursor()
    if request.method == 'PATCH':
        d = request.json or {}; accion = d.get('accion')
        if accion == 'registrar_fisico':
            # Validar stock_fisico finito (NaN/Inf rompe kardex)
            try:
                sf = float(d.get('stock_fisico', 0))
                import math as _m
                if not _m.isfinite(sf) or sf < 0:
                    return jsonify({'error': 'stock_fisico inválido'}), 400
            except (TypeError, ValueError):
                return jsonify({'error': 'stock_fisico no numérico'}), 400
            c.execute("""UPDATE conteo_items SET stock_fisico=?,diferencia=?-stock_sistema,observaciones=?
                         WHERE conteo_id=? AND codigo_mp=?""",
                      (sf, sf, d.get('observaciones',''), cid, d.get('codigo_mp')))
            c.execute("""UPDATE conteos_fisicos SET
                         items_diferencia=(SELECT COUNT(*) FROM conteo_items
                                          WHERE conteo_id=? AND stock_fisico IS NOT NULL AND ABS(diferencia)>0.1)
                         WHERE id=?""", (cid, cid))
        elif accion == 'cerrar':
            c.execute("UPDATE conteos_fisicos SET estado='Cerrado',fecha_cierre=?,aprobado_por=? WHERE id=?",
                      (datetime.now().isoformat(), d.get('aprobado_por',''), cid))
        elif accion == 'aplicar_ajustes':
            # Solo auto-ajustar diferencias que NO requieren gerencia · las
            # >5% se escalan (BDG-PRO-002) · sin este filtro esta ruta legacy
            # saltaba el control que conteo_cerrar/conteo_ajustar sí respetan.
            # FIX P0 audit 24-may-2026 · preservar lote real cuando exista
            # en conteo_items.lote (la ruta conteo_cerrar:7676 ya lo hace).
            # Antes esta ruta legacy escribía siempre 'AJUSTE-CICLICO-<id>'
            # aún si había lote real → pérdida de trazabilidad INVIMA.
            # B1 (12-jun): SELECT incluye id + atomic claim por FILA antes de insertar.
            # Antes marcaba ajuste_aplicado por codigo (no por id/lote) y sin claim ->
            # 2 PATCH concurrentes multi-worker podian doble-insertar (M20/M3). Ahora
            # solo inserta si CLAIMA la fila (rowcount=1); idempotente y por-lote.
            c.execute("SELECT id,codigo_mp,nombre_mp,diferencia,COALESCE(lote,'') FROM conteo_items "
                      "WHERE conteo_id=? AND ABS(diferencia)>0.1 AND ajuste_aplicado=0 "
                      "AND COALESCE(requiere_gerencia,0)=0", (cid,))
            _resp = d.get('responsable','') or session.get('compras_user','')
            for _it_id, cod, nom, dif, lote_real in c.fetchall():
                c.execute("UPDATE conteo_items SET ajuste_aplicado=1 "
                          "WHERE id=? AND COALESCE(ajuste_aplicado,0)=0", (_it_id,))
                if c.rowcount == 0:
                    continue  # otra request ya aplico esta fila (idempotente)
                tipo = 'Entrada' if dif > 0 else 'Salida'
                lote_obj = lote_real or f'AJUSTE-CICLICO-{cid}'
                c.execute("""INSERT INTO movimientos (material_id,material_nombre,cantidad,tipo,fecha,observaciones,lote,estado_lote,operador)
                             VALUES (?,?,?,?,?,?,?,'VIGENTE',?)""",
                          (cod, nom, abs(dif), tipo, datetime.now().isoformat(),
                           f'Ajuste conteo {cid}' + (f' · lote {lote_real}' if lote_real else ''),
                           lote_obj, _resp))
                c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                             VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                          (_resp, 'AJUSTE_INVENTARIO_CONTEO', 'conteo_items', str(cid),
                           f'MP:{cod} {tipo} {abs(dif)}g · conteo #{cid}' + (f' · lote {lote_real}' if lote_real else ''),
                           request.remote_addr if request else ''))
        conn.commit()
        c.execute("SELECT id,numero,estado,total_items,items_diferencia FROM conteos_fisicos WHERE id=?", (cid,))
        r = c.fetchone()
        return jsonify({'id':r[0],'numero':r[1],'estado':r[2],'total':r[3],'diffs':r[4]})
    c.execute("SELECT * FROM conteos_fisicos WHERE id=?", (cid,)); h = c.fetchone()
    if not h: return jsonify({'error':'No encontrado'}), 404
    c.execute("SELECT codigo_mp,nombre_mp,stock_sistema,stock_fisico,diferencia,ajuste_aplicado,observaciones FROM conteo_items WHERE conteo_id=? ORDER BY nombre_mp", (cid,))
    items = [{'codigo':r[0],'nombre':r[1],'sistema':r[2],'fisico':r[3],'diff':r[4],'ajustado':r[5],'obs':r[6]} for r in c.fetchall()]
    return jsonify({'header':{'id':h[0],'numero':h[1],'estado':h[4],'responsable':h[5],'total':h[7],'diffs':h[8]},'items':items})

@bp.route('/api/lotes/cuarentena/<int:mov_id>/liberar', methods=['POST'])
def liberar_cuarentena(mov_id):
    # Liberación de cuarentena = decisión de Calidad. Antes solo admins,
    # ahora QC + admins (consistente con cc-review y liberar_lote).
    u, err, code = _require_qc()
    if err:
        return err, code
    d = request.json or {}
    # INVIMA-FIX · 21-may-2026 · decision whitelist (antes cualquier string)
    decision = d.get('decision','Aprobado')
    if decision not in ('Aprobado', 'Rechazado'):
        return jsonify({'error': "decision debe ser 'Aprobado' o 'Rechazado'"}), 400
    # INVIMA-FIX · 21-may-2026 · filtrar estado actual (no revivir RECHAZADO)
    conn = get_db(); c = conn.cursor()
    estado_actual_row = c.execute(
        "SELECT estado_lote FROM movimientos WHERE id=?", (mov_id,),
    ).fetchone()
    if not estado_actual_row:
        return jsonify({'error': 'Movimiento no existe'}), 404
    estado_actual = estado_actual_row[0] or ''
    if estado_actual not in ('CUARENTENA', 'CUARENTENA_EXTENDIDA'):
        return jsonify({
            'error': f'No se puede liberar lote en estado {estado_actual}',
            'codigo': 'ESTADO_NO_LIBERABLE',
        }), 409
    # Firma electrónica Part 11 §11.200 · misma exigencia que liberar_lote /
    # cc-review · si no se gateara aquí, el control sería eludible por esta ruta.
    meaning = 'libera' if decision == 'Aprobado' else 'rechaza'
    sig_id = d.get('signature_id')
    if not _validar_e_sign(c, sig_id, record_table='movimientos', record_id=mov_id,
                           meaning=meaning, signer_username=u):
        return jsonify({
            'error': 'Firma electrónica requerida',
            'requiere_firma': True,
            'sign_meaning': meaning,
            'record_table': 'movimientos',
            'record_id': str(mov_id),
            'detail': f"Firmá vía POST /api/sign con meaning='{meaning}', "
                      f"record_table='movimientos', record_id='{mov_id}' y reenviá signature_id.",
        }), 400
    nuevo_estado = 'VIGENTE' if decision == 'Aprobado' else 'RECHAZADO'
    c.execute("UPDATE movimientos SET estado_lote=? WHERE id=?", (nuevo_estado, mov_id))
    c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                 VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
              (u, f'{decision.upper()}_CUARENTENA', 'movimientos',
               str(mov_id), (d.get('observaciones','') + f' · e-sign #{sig_id}'), request.remote_addr))
    conn.commit()
    return jsonify({'ok':True, 'decision':decision, 'estado':nuevo_estado, 'signature_id': sig_id})


@bp.route('/api/lotes/cuarentena/liberar-inventario', methods=['POST'])
def liberar_cuarentena_inventario():
    """Sebastián 16-jun · día de inventario · liberación rápida a inventario.

    Saca de cuarentena a VIGENTE la MP recibida, SIN la firma Part 11 por lote
    (fricción que no aplica el día de inventario). Es la contraparte del interruptor
    RECEPCION_AUTO_VIGENTE: solo funciona mientras ese interruptor está ENCENDIDO y
    solo para ADMIN. Queda 100% auditado (audit_log por lote). Al apagar el
    interruptor tras el inventario, esta ruta se cierra y vuelve la liberación
    formal con firma (cc-review / liberar_cuarentena).

    Body: {mov_id?} · si se da mov_id libera ese lote; si no, libera TODOS los
    lotes en CUARENTENA / CUARENTENA_EXTENDIDA.
    """
    u, err, code = _require_admin()
    if err:
        return err, code
    d = request.json or {}
    mov_id = d.get('mov_id')
    conn = get_db(); c = conn.cursor()
    from database import recepcion_auto_vigente as _rav
    if not _rav(c):
        return jsonify({
            'error': 'Liberación rápida deshabilitada (el modo inventario está apagado). '
                     'Usá la liberación formal con firma desde Calidad → Cuarentena.',
            'codigo': 'SWITCH_OFF',
        }), 409
    if mov_id:
        filas = c.execute(
            "SELECT id, material_id, lote FROM movimientos WHERE id=? "
            "AND UPPER(COALESCE(estado_lote,'')) IN ('CUARENTENA','CUARENTENA_EXTENDIDA')",
            (mov_id,)).fetchall()
    else:
        filas = c.execute(
            "SELECT id, material_id, lote FROM movimientos "
            "WHERE UPPER(COALESCE(estado_lote,'')) IN ('CUARENTENA','CUARENTENA_EXTENDIDA')"
        ).fetchall()
    liberados = []
    for fid, fmat, flote in filas:
        c.execute("UPDATE movimientos SET estado_lote='VIGENTE' WHERE id=?", (fid,))
        try:
            c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                         VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                      (u, 'LIBERAR_CUARENTENA_INVENTARIO', 'movimientos', str(fid),
                       f'Liberación rápida día inventario · {fmat} lote {flote} → VIGENTE',
                       request.remote_addr if request else ''))
        except sqlite3.OperationalError:
            pass
        liberados.append({'mov_id': fid, 'material_id': fmat, 'lote': flote})
    conn.commit()
    return jsonify({'ok': True, 'liberados': len(liberados), 'detalle': liberados})


@bp.route('/api/inventario/modo-inventario', methods=['GET', 'POST'])
def modo_inventario_config():
    """Sebastián 16-jun · toggle 'modo inventario' = recepciones entran directo a
    inventario (VIGENTE) sin pasar por cuarentena de Calidad. Guardado en
    app_settings (botón en la UI · ADMIN · sin tocar Render · efecto inmediato y
    reversible). Default OFF = posición INVIMA. GET → estado; POST {activo:bool}."""
    conn = get_db(); c = conn.cursor()
    try:
        c.execute("""CREATE TABLE IF NOT EXISTS app_settings (
            clave TEXT PRIMARY KEY, valor TEXT NOT NULL, descripcion TEXT,
            actualizado_at_utc TEXT, actualizado_por TEXT, tenant_id INTEGER DEFAULT 1)""")
    except Exception:
        pass
    if request.method == 'POST':
        u, err, code = _require_admin()
        if err:
            return err, code
        body = request.get_json(silent=True) or {}
        activo = bool(body.get('activo'))
        val = '1' if activo else '0'
        c.execute(
            "INSERT INTO app_settings (clave,valor,descripcion,actualizado_at_utc,actualizado_por) "
            "VALUES ('recepcion_auto_vigente',?,?,datetime('now'),?) "
            "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor, "
            "actualizado_at_utc=excluded.actualizado_at_utc, actualizado_por=excluded.actualizado_por",
            (val, 'Recepción entra directo a inventario sin cuarentena (modo inventario)', u))
        try:
            audit_log(c, usuario=u, accion='SET_MODO_INVENTARIO', tabla='app_settings',
                      registro_id='recepcion_auto_vigente', despues={'activo': activo})
        except Exception:
            pass
        conn.commit()
        return jsonify({'ok': True, 'activo': activo})
    # GET
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    from database import recepcion_auto_vigente as _rav
    return jsonify({'ok': True, 'activo': bool(_rav(c))})


@bp.route('/api/inventario/reset-inventario-cero', methods=['GET', 'POST'])
def reset_inventario_cero():
    """Sebastián 16-jun · app EN PRUEBA · pone TODO el stock en CERO conservando el
    catálogo (maestro_mps con códigos/INCI/nombres, fórmulas, mapeos). Borra el kardex
    de MP (movimientos) y, si se pide, el de envases (movimientos_mee + maestro_mee.
    stock_actual=0) y los conteos viejos. Hace RESPALDO (tablas *_bak_<ts>) antes de
    borrar. ADMIN · GET=preview · POST {confirmar:'CERO', incluir_mee, limpiar_conteos}.
    NO toca maestro_mps/maestro_mee (filas), fórmulas, sku maps ni stock_pt."""
    u, err, code = _require_admin()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()

    def _count(q):
        try:
            return c.execute(q).fetchone()[0]
        except Exception:
            return 0
    preview = {
        'movimientos_mp': _count("SELECT COUNT(*) FROM movimientos"),
        'mps_con_saldo': _count("SELECT COUNT(DISTINCT material_id) FROM movimientos"),
        'movimientos_mee': _count("SELECT COUNT(*) FROM movimientos_mee"),
        'conteos': _count("SELECT COUNT(*) FROM conteos_fisicos"),
        'catalogo_mps_intacto': _count("SELECT COUNT(*) FROM maestro_mps"),
    }
    if request.method == 'GET' or not (request.get_json(silent=True) or {}).get('confirmar'):
        return jsonify({'ok': True, 'dry_run': True, **preview,
                        'nota': "POST con {confirmar:'CERO'} para ejecutar"})

    body = request.get_json(silent=True) or {}
    if (body.get('confirmar') or '').strip().upper() != 'CERO':
        return jsonify({'ok': False, 'error': "Para ejecutar enviá confirmar='CERO'"}), 400
    incluir_mee = bool(body.get('incluir_mee', True))
    limpiar_conteos = bool(body.get('limpiar_conteos', True))

    import datetime as _dt_r
    ts = (_dt_r.datetime.utcnow() - _dt_r.timedelta(hours=5)).strftime('%Y%m%d_%H%M%S')
    backups = []
    # RESPALDO antes de borrar (CREATE TABLE AS SELECT · SQLite + PG)
    try:
        c.execute(f"CREATE TABLE movimientos_bak_{ts} AS SELECT * FROM movimientos")
        backups.append(f"movimientos_bak_{ts}")
    except Exception as _e:
        log.warning("reset: backup movimientos falló: %s", _e)

    mp_borrados = preview['movimientos_mp']
    c.execute("DELETE FROM movimientos")

    mee_borrados = 0
    if incluir_mee:
        try:
            c.execute(f"CREATE TABLE movimientos_mee_bak_{ts} AS SELECT * FROM movimientos_mee")
            backups.append(f"movimientos_mee_bak_{ts}")
        except Exception:
            pass
        mee_borrados = preview['movimientos_mee']
        c.execute("DELETE FROM movimientos_mee")
        try:
            c.execute("UPDATE maestro_mee SET stock_actual=0")
        except Exception:
            pass

    conteos_borrados = 0
    if limpiar_conteos:
        conteos_borrados = preview['conteos']
        try:
            c.execute("DELETE FROM conteo_items")
            c.execute("DELETE FROM conteos_fisicos")
        except Exception:
            pass

    try:
        audit_log(c, usuario=u, accion='RESET_INVENTARIO_CERO', tabla='movimientos',
                  registro_id=0, despues={'mp_borrados': mp_borrados, 'mee_borrados': mee_borrados,
                                          'conteos_borrados': conteos_borrados, 'backups': backups})
    except Exception:
        pass
    conn.commit()
    return jsonify({'ok': True, 'dry_run': False, 'mp_borrados': mp_borrados,
                    'mee_borrados': mee_borrados, 'conteos_borrados': conteos_borrados,
                    'catalogo_conservado': preview['catalogo_mps_intacto'], 'backups': backups})


@bp.route('/admin/reset-inventario', methods=['GET'])
def reset_inventario_page():
    """Página ADMIN para poner el inventario en CERO conservando el catálogo (app en
    prueba). Muestra preview + confirmación tipeada antes de ejecutar."""
    u, err, code = _require_admin()
    if err:
        return err, code
    from flask import Response
    html = (
        '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        '<title>Reset inventario a cero</title>'
        '<style>body{font-family:system-ui,Segoe UI,Roboto,sans-serif;margin:0;background:#f8fafc;color:#1e293b}'
        '.wrap{max-width:720px;margin:0 auto;padding:24px}h1{font-size:21px;margin:0 0 6px}'
        '.warn{background:#fef2f2;border:1px solid #fecaca;color:#991b1b;border-radius:10px;padding:14px 18px;margin:14px 0;font-size:14px}'
        '.card{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:16px 18px;margin:14px 0}'
        '.k{font-size:26px;font-weight:800}label{display:block;margin:10px 0;font-size:14px}'
        'input[type=text]{padding:9px 12px;border:1.5px solid #cbd5e1;border-radius:8px;font-size:15px;width:160px}'
        'button{border:none;border-radius:9px;padding:11px 20px;font-weight:800;font-size:15px;cursor:pointer}'
        '.go{background:#b91c1c;color:#fff}.go:disabled{opacity:.4;cursor:not-allowed}'
        'a.back{color:#0d9488;text-decoration:none;font-size:13px}</style></head><body><div class="wrap">'
        '<a class="back" href="/inventarios">← Volver a Inventarios</a>'
        '<h1>🧹 Reset de inventario a CERO</h1>'
        '<div class="warn"><b>Solo para la app en prueba.</b> Pone TODO el stock en cero para cargar el conteo nuevo. '
        '<b>Conserva</b> el catálogo (códigos, INCI, nombres), fórmulas y mapeos. <b>Borra</b> el kardex de MP (y envases) — se hace un respaldo antes.</div>'
        '<div class="card" id="preview">Cargando preview…</div>'
        '<div class="card">'
        '<label><input type="checkbox" id="mee" checked> También poner envases/MEE en cero</label>'
        '<label><input type="checkbox" id="conteos" checked> Borrar conteos físicos viejos</label>'
        '<label>Escribí <b>CERO</b> para confirmar: <input type="text" id="confirmar" autocomplete="off" oninput="document.getElementById(\'go\').disabled = this.value.trim().toUpperCase() !== \'CERO\'"></label>'
        '<button class="go" id="go" disabled onclick="ejecutar()">🧹 Poner inventario en CERO</button>'
        '</div><div id="msg" style="font-size:14px;margin-top:10px"></div>'
        '<script>'
        'function _csrf(){return document.cookie.split(";").find(function(c){return c.trim().indexOf("csrf_token=")===0})?.split("=")[1]||"";}'
        'async function cargar(){'
        'try{var r=await fetch("/api/inventario/reset-inventario-cero");var d=await r.json();'
        'document.getElementById("preview").innerHTML="<div class=k>"+d.movimientos_mp+"</div>movimientos de MP a borrar · "+d.mps_con_saldo+" MP con saldo<br>"+d.movimientos_mee+" movimientos de envases · "+d.conteos+" conteos<br><span style=color:#166534>✓ catálogo intacto: "+d.catalogo_mps_intacto+" MP (códigos/INCI/nombres se conservan)</span>";'
        '}catch(e){document.getElementById("preview").textContent="No se pudo cargar preview: "+e;}}'
        'async function ejecutar(){'
        'if(!confirm("¿Seguro? Esto pone TODO el stock en cero (con respaldo). Conserva el catálogo.")) return;'
        'var b=document.getElementById("go");b.disabled=true;b.textContent="Ejecutando…";'
        'try{var r=await fetch("/api/inventario/reset-inventario-cero",{method:"POST",headers:{"Content-Type":"application/json","X-CSRFToken":_csrf()},body:JSON.stringify({confirmar:"CERO",incluir_mee:document.getElementById("mee").checked,limpiar_conteos:document.getElementById("conteos").checked})});'
        'var d=await r.json();'
        'if(!r.ok||!d.ok){document.getElementById("msg").innerHTML="<span style=color:#991b1b>Error: "+(d.error||r.status)+"</span>";b.disabled=false;b.textContent="🧹 Poner inventario en CERO";return;}'
        'document.getElementById("msg").innerHTML="<span style=color:#166534;font-weight:700>✅ Listo. Stock en cero. MP borradas: "+d.mp_borrados+" · envases: "+d.mee_borrados+" · conteos: "+d.conteos_borrados+" · catálogo conservado: "+d.catalogo_conservado+". Respaldo: "+(d.backups||[]).join(", ")+"</span><br>Ya podés cargar el conteo nuevo desde Inventarios › Ingreso MP.";'
        'b.textContent="✓ Hecho";cargar();'
        '}catch(e){document.getElementById("msg").textContent="Error: "+e;b.disabled=false;b.textContent="🧹 Poner inventario en CERO";}}'
        'cargar();'
        '</script></div></body></html>')
    return Response(html, mimetype="text/html")


@bp.route('/api/maestro-mp/<codigo>/precio', methods=['POST'])
def actualizar_precio_mp(codigo):
    """Actualiza precio_referencia de una MP · auditado.

    Solo Compras/Calidad/Admin · cambios de precio impactan margen y
    presupuestos, requieren trazabilidad.
    """
    u, err, code_err = _require_planta_write()
    if err:
        return err, code_err
    d = request.json or {}
    # Validar precio (allow_zero para "limpiar" precio temporalmente)
    precio, perr = validate_money(d.get('precio_kg', 0), allow_zero=True,
                                   field_name='precio_kg')
    if perr:
        return jsonify(perr), 400
    conn = get_db(); c = conn.cursor()
    antes_row = c.execute(
        "SELECT codigo_mp, nombre_comercial, precio_referencia, ultima_act_precio "
        "FROM maestro_mps WHERE codigo_mp=?", (codigo,)).fetchone()
    if not antes_row:
        return jsonify({'error': 'MP no encontrada'}), 404
    antes = dict(antes_row)
    c.execute("UPDATE maestro_mps SET precio_referencia=?,ultima_act_precio=? WHERE codigo_mp=?",
              (precio, datetime.now().isoformat()[:10], codigo))
    c.execute("""INSERT INTO precios_mp_historico (codigo_mp,proveedor,precio_kg,fecha,origen,observaciones)
                 VALUES (?,?,?,?,?,?)""",
              (codigo, d.get('proveedor',''), precio, datetime.now().isoformat()[:10],
               d.get('origen','manual'), d.get('observaciones','')))
    audit_log(c, usuario=u, accion='ACTUALIZAR_PRECIO_MP', tabla='maestro_mps',
              registro_id=codigo, antes=antes,
              despues={'precio_referencia': precio, 'proveedor': d.get('proveedor',''),
                       'origen': d.get('origen','manual')},
              detalle=f"Actualizó precio MP {codigo}: "
                      f"{antes.get('precio_referencia') or 0} → {precio}"
                      + (f" · {d.get('observaciones','')}" if d.get('observaciones') else ""))
    conn.commit()
    return jsonify({'ok':True, 'precio_kg':precio})

@bp.route('/api/admin/backfill-precios-mp', methods=['POST'])
def backfill_precios_mp():
    """Pobla precio_referencia en maestro_mps desde movimientos.precio_kg y precios_mp_historico.
    Solo actualiza MPs que tienen precio_referencia=0 o nulo."""
    # P0 audit 26-may-2026 · admin gate · este endpoint hacía UPDATE masivo
    # sobre maestro_mps.precio_referencia (data financiera para sugerencias
    # de OC) sin ningún chequeo de sesión ni rol.
    from config import ADMIN_USERS as _AU
    _user = session.get('compras_user', '')
    if (_user or '').lower() not in {x.lower() for x in _AU}:
        return jsonify({'error': 'Solo admin'}), 403
    conn = get_db(); c = conn.cursor()
    actualizados = 0
    # Fuente 1: promedio ponderado de movimientos de entrada con precio registrado
    c.execute("""SELECT material_id, AVG(precio_kg) as avg_precio
                 FROM movimientos
                 WHERE tipo='Entrada' AND precio_kg IS NOT NULL AND precio_kg > 0
                 GROUP BY material_id""")
    from_movs = c.fetchall()
    for mat_id, avg_p in from_movs:
        c.execute("""UPDATE maestro_mps SET precio_referencia=?, ultima_act_precio=datetime('now', '-5 hours')
                     WHERE codigo_mp=? AND (precio_referencia IS NULL OR precio_referencia=0)""",
                  (round(avg_p, 2), mat_id))
        actualizados += c.rowcount
    # Fuente 2: precios_mp_historico (precio más reciente por MP)
    c.execute("""SELECT codigo_mp, precio_kg FROM precios_mp_historico
                 WHERE (codigo_mp, fecha) IN (
                     SELECT codigo_mp, MAX(fecha) FROM precios_mp_historico
                     WHERE precio_kg > 0 GROUP BY codigo_mp
                 )""")
    from_hist = c.fetchall()
    hist_count = 0
    for codigo, precio in from_hist:
        c.execute("""UPDATE maestro_mps SET precio_referencia=?, ultima_act_precio=datetime('now', '-5 hours')
                     WHERE codigo_mp=? AND (precio_referencia IS NULL OR precio_referencia=0)""",
                  (round(precio, 2), codigo))
        hist_count += c.rowcount
    actualizados += hist_count
    # Stats
    c.execute("SELECT COUNT(*) FROM maestro_mps WHERE precio_referencia > 0")
    con_precio = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM maestro_mps WHERE activo=1")
    total_activos = c.fetchone()[0]
    # Audit log · mutación masiva de precios financieros
    try:
        from audit_helpers import audit_log as _al
        _al(c, usuario=_user, accion='BACKFILL_PRECIOS_MP', tabla='maestro_mps',
            despues={'actualizados': actualizados, 'hist_count': hist_count,
                     'con_precio_ahora': con_precio, 'total_activos': total_activos},
            detalle=f'Backfill precio_referencia · {actualizados} MPs actualizados')
    except Exception as _ae:
        import logging as _lg
        _lg.getLogger('inventario').warning('audit backfill_precios_mp fallo: %s', _ae)
    conn.commit()
    return jsonify({
        'ok': True,
        'actualizados': actualizados,
        'con_precio_ahora': con_precio,
        'total_activos': total_activos,
        'cobertura_pct': round(con_precio / total_activos * 100, 1) if total_activos > 0 else 0
    })

# ═══════════════════════════════════════════════════════════════════════
# MÓDULO MEE — Material de Empaque y Envase  (Tasks #70 #71 #72)
# ═══════════════════════════════════════════════════════════════════════

def _init_mee_movimientos():
    conn = get_db(); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS movimientos_mee (
        id          INTEGER  PRIMARY KEY AUTOINCREMENT,
        mee_codigo  TEXT     NOT NULL,
        tipo        TEXT     NOT NULL CHECK(tipo IN ('Entrada','Salida','Ajuste')),
        cantidad    REAL     NOT NULL,
        unidad      TEXT     DEFAULT 'und',
        lote_ref    TEXT     DEFAULT '',
        batch_ref   TEXT     DEFAULT '',
        responsable TEXT     DEFAULT '',
        observaciones TEXT   DEFAULT '',
        fecha       DATETIME DEFAULT (datetime('now', '-5 hours')),
        anulado     INTEGER  DEFAULT 0
    )""")
    conn.commit()

_init_mee_movimientos()

@bp.route('/api/mee', methods=['POST'])
def mee_crear():
    """Crea un nuevo material en maestro_mee."""
    # P0 audit 26-may-2026 · sin gate, cualquier anónimo podía crear MEE.
    _u, _err, _code = _require_planta_write()
    if _err:
        return _err, _code
    d = request.json or {}
    codigo      = d.get('codigo','').strip().upper()
    descripcion = d.get('descripcion','').strip()
    categoria   = d.get('categoria','Otro').strip()
    proveedor   = d.get('proveedor','').strip()
    unidad      = d.get('unidad','und').strip()
    # FIX P1 audit 24-may-2026 · antes float() pelado aceptaba NaN/Infinity
    # /negativos. Ahora validate_money con allow_zero (stock=0 al crear es
    # legítimo). max_value=10M unidades (cap razonable para inventario).
    from http_helpers import validate_money
    stock_actual, err = validate_money(d.get('stock_actual', 0),
                                        allow_zero=True, max_value=10_000_000,
                                        field_name='stock_actual')
    if err:
        return jsonify(err), 400
    stock_minimo, err = validate_money(d.get('stock_minimo', 0),
                                        allow_zero=True, max_value=10_000_000,
                                        field_name='stock_minimo')
    if err:
        return jsonify(err), 400
    if not codigo or not descripcion:
        return jsonify({'error': 'codigo y descripcion requeridos'}), 400
    conn = get_db(); c = conn.cursor()
    try:
        c.execute("""INSERT INTO maestro_mee (codigo, descripcion, categoria, unidad, proveedor,
                                              stock_actual, stock_minimo, estado)
                     VALUES (?,?,?,?,?,?,?,'Activo')""",
                  (codigo, descripcion, categoria, unidad, proveedor, stock_actual, stock_minimo))
        # Audit log creación MEE
        try:
            from audit_helpers import audit_log as _al
            _al(c, usuario=_u, accion='CREAR_MEE', tabla='maestro_mee', registro_id=codigo,
                despues={'codigo': codigo, 'descripcion': descripcion[:80],
                         'categoria': categoria, 'proveedor': proveedor[:80],
                         'stock_actual': stock_actual, 'stock_minimo': stock_minimo},
                detalle=f'Material MEE {codigo} creado')
        except Exception as _ae:
            import logging as _lg
            _lg.getLogger('inventario').warning('audit crear_mee fallo: %s', _ae)
        conn.commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    return jsonify({'ok': True, 'codigo': codigo, 'message': f'Material {codigo} creado exitosamente'})

@bp.route('/api/mee/stock', methods=['GET'])
def mee_stock_list():
    """Lista maestro_mee con stock, alertas y metricas de movimiento."""
    conn = get_db(); c = conn.cursor()
    cat_f = request.args.get('categoria', '')
    # FIX 12-jun · stock CANÓNICO = SUM(movimientos_mee) (igual que _mee_stock_real),
    # no el cache m.stock_actual que driftea (hay backfill de drift en admin). El
    # display y la alerta critico/bajo se calculan sobre stock_real, no el cache.
    sql = """
        SELECT m.codigo, m.descripcion, m.categoria, m.unidad,
               COALESCE(mv.stock_real, m.stock_actual, 0) as stock_actual, m.stock_minimo, m.estado, m.proveedor,
               COALESCE(mv.ultima_entrada,'') as ultima_entrada,
               COALESCE(mv.ultima_salida,'')  as ultima_salida,
               COALESCE(mv.total_entradas,0)  as total_entradas,
               COALESCE(mv.total_salidas,0)   as total_salidas
        FROM maestro_mee m
        LEFT JOIN (
            SELECT mee_codigo,
                   MAX(CASE WHEN tipo='Entrada' AND anulado=0 THEN fecha END) as ultima_entrada,
                   MAX(CASE WHEN tipo='Salida'  AND anulado=0 THEN fecha END) as ultima_salida,
                   SUM(CASE WHEN tipo='Entrada' AND anulado=0 THEN cantidad ELSE 0 END) as total_entradas,
                   SUM(CASE WHEN tipo='Salida'  AND anulado=0 THEN cantidad ELSE 0 END) as total_salidas,
                   SUM(CASE WHEN anulado=0 AND tipo='Entrada' THEN cantidad
                            WHEN anulado=0 AND tipo='Salida'  THEN -cantidad
                            WHEN anulado=0 AND tipo='Ajuste'  THEN cantidad
                            ELSE 0 END) as stock_real
            FROM movimientos_mee GROUP BY mee_codigo
        ) mv ON m.codigo = mv.mee_codigo
        WHERE m.estado='Activo'
    """
    params = []
    if cat_f:
        sql += " AND m.categoria=?"; params.append(cat_f)
    sql += " ORDER BY m.categoria, m.descripcion"
    c.execute(sql, params)
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]

    from datetime import date
    hoy = date.today()
    for r in rows:
        # canónico clamp >=0 (igual que _mee_stock_real) · el display y la alerta
        # usan el stock real, nunca negativo aunque haya sobre-consumo registrado
        s = max(float(r['stock_actual'] or 0), 0)
        r['stock_actual'] = s
        mn = r['stock_minimo'] or 0
        if mn > 0:
            ratio = s / mn
            if ratio <= 0:    r['alerta'] = 'critico'
            elif ratio < 1:   r['alerta'] = 'bajo'
            elif ratio < 1.5: r['alerta'] = 'advertencia'
            else:             r['alerta'] = 'ok'
        else:
            r['alerta'] = 'sin_minimo'
        ref = r['ultima_entrada'] or r['ultima_salida']
        if ref:
            try:
                days = (hoy - date.fromisoformat(ref[:10])).days
                r['dias_sin_mov'] = days
                r['obsoleto'] = days > 90
            except Exception:
                r['dias_sin_mov'] = None; r['obsoleto'] = False
        else:
            r['dias_sin_mov'] = None; r['obsoleto'] = s > 0

    c.execute("SELECT DISTINCT categoria FROM maestro_mee WHERE estado='Activo' ORDER BY categoria")
    categorias = [row[0] for row in c.fetchall()]
    c.execute("SELECT COUNT(*) FROM maestro_mee WHERE estado='Activo'")
    total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM maestro_mee WHERE estado='Activo' AND stock_minimo>0 AND stock_actual<stock_minimo")
    bajo = c.fetchone()[0]
    return jsonify({'items': rows, 'categorias': categorias, 'total': total, 'bajo_minimo': bajo})

@bp.route('/api/mee/movimiento', methods=['POST'])
def mee_registrar_movimiento():
    """Registra una entrada, salida o ajuste de empaque MEE."""
    # P0 audit 26-may-2026 · sin gate · cualquier anónimo movía stock MEE.
    _u, _err, _code = _require_planta_write()
    if _err:
        return _err, _code
    d = request.json or {}
    codigo      = d.get('codigo','').strip()
    tipo        = d.get('tipo','').strip()
    # FIX P1 audit 24-may-2026 · validar cantidad numérica (no NaN/Inf/abs).
    # Para Ajuste, allow_zero permite registrar stock_objetivo=0 (vaciar).
    from http_helpers import validate_money
    cantidad, err = validate_money(d.get('cantidad', 0),
                                    allow_zero=(tipo == 'Ajuste'),
                                    max_value=10_000_000,
                                    field_name='cantidad')
    if err:
        return jsonify(err), 400
    unidad      = d.get('unidad','und').strip()
    lote_ref    = d.get('lote_ref','').strip()
    batch_ref   = d.get('batch_ref','').strip()
    responsable = d.get('responsable', session.get('compras_user','')).strip()
    obs         = d.get('observaciones','').strip()

    if not codigo or tipo not in ('Entrada','Salida','Ajuste'):
        return jsonify({'error': 'codigo, tipo (Entrada/Salida/Ajuste) requeridos'}), 400
    if tipo != 'Ajuste' and cantidad <= 0:
        return jsonify({'error': 'cantidad debe ser > 0 para Entrada/Salida'}), 400

    conn = get_db(); c = conn.cursor()
    c.execute("SELECT codigo, descripcion, stock_actual, unidad FROM maestro_mee WHERE codigo=?", (codigo,))
    mee = c.fetchone()
    if not mee:
        return jsonify({'error': f'Codigo MEE {codigo} no encontrado'}), 404
    stock_ant = float(mee[2] or 0)

    # Fix 28-may · rechazar Salida que excede el stock (como el pre-check de MP)
    # en vez de clampar a 0 registrando la cantidad completa → eso dejaba drift
    # entre maestro_mee.stock_actual (0) y SUM(movimientos_mee). Para corregir
    # stock a la baja sin venta real, usar 'Ajuste' (stock objetivo).
    if tipo == 'Salida' and cantidad > stock_ant + 0.001:
        return jsonify({
            'error': 'stock insuficiente',
            'stock_actual': stock_ant,
            'cantidad_pedida': cantidad,
            'sugerencia': 'Usar Ajuste para fijar el stock objetivo',
        }), 422

    # Para 'Ajuste', 'cantidad' es el stock OBJETIVO (absoluto). El movimiento
    # debe registrar el DELTA (objetivo - actual) · si registrara el absoluto,
    # SUM(movimientos_mee) driftearía contra stock_actual. Entrada/Salida sí
    # registran su propia magnitud (ya son deltas).
    mov_cantidad = round(cantidad - stock_ant, 2) if tipo == 'Ajuste' else cantidad

    c.execute("""INSERT INTO movimientos_mee
                 (mee_codigo, tipo, cantidad, unidad, lote_ref, batch_ref, responsable, observaciones)
                 VALUES (?,?,?,?,?,?,?,?)""",
              (codigo, tipo, mov_cantidad, unidad, lote_ref, batch_ref, responsable, obs))
    mov_id = c.lastrowid

    if tipo == 'Entrada':
        c.execute("UPDATE maestro_mee SET stock_actual = stock_actual + ? WHERE codigo=?", (cantidad, codigo))
    elif tipo == 'Salida':
        c.execute("UPDATE maestro_mee SET stock_actual = MAX(0, stock_actual - ?) WHERE codigo=?", (cantidad, codigo))
    else:  # Ajuste · cantidad = stock objetivo absoluto
        c.execute("UPDATE maestro_mee SET stock_actual = ? WHERE codigo=?", (cantidad, codigo))

    c.execute("SELECT stock_actual, stock_minimo FROM maestro_mee WHERE codigo=?", (codigo,))
    s_new, s_min = c.fetchone()
    try:
        c.execute("""INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
                     VALUES (?,?,?,?,?,?,datetime('now', '-5 hours'))""",
                  (responsable, 'MOVIMIENTO_MEE', 'movimientos_mee', str(mov_id),
                   f'{tipo} {mov_cantidad} de {codigo} · stock {stock_ant}->{s_new}',
                   request.remote_addr if request else ''))
    except sqlite3.OperationalError:
        pass
    conn.commit()

    alerta = None
    if s_min and s_min > 0 and s_new < s_min:
        alerta = f'Stock bajo minimo: {s_new:.0f} {unidad} (minimo: {s_min:.0f})'
    return jsonify({
        'ok': True, 'movimiento_id': mov_id, 'stock_nuevo': s_new, 'alerta': alerta,
        'message': f'{tipo} de {cantidad:.0f} {unidad} registrada para {mee[1]}'
    })

@bp.route('/api/mee/movimientos', methods=['GET'])
def mee_historial_movimientos():
    """Historial paginado de movimientos MEE. Sprint MEE PRO 20-may-2026.

    Query params:
      codigo · filtra por código MEE específico
      tipo · Entrada/Salida/Ajuste
      q · busca en mee_codigo / descripcion / lote_ref / batch_ref / obs
      limit · default 50, max 500
      offset · default 0
      incluir_anulados · 1 para mostrar anulados
    """
    codigo = (request.args.get('codigo','') or '').strip()
    tipo   = (request.args.get('tipo','') or '').strip()
    q      = (request.args.get('q','') or '').strip().lower()
    try:
        limit = max(1, min(int(request.args.get('limit', 50)), 500))
    except (ValueError, TypeError):
        limit = 50
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    incluir_anul = (request.args.get('incluir_anulados','') or '').strip() in ('1','true','yes')

    where = ['1=1' if incluir_anul else 'mv.anulado=0']
    params = []
    if codigo:
        where.append("mv.mee_codigo=?"); params.append(codigo)
    if tipo:
        where.append("mv.tipo=?"); params.append(tipo)
    if q:
        q_safe = q.replace('\\','\\\\').replace('%','\\%').replace('_','\\_')
        like = f'%{q_safe}%'
        where.append(
            "(LOWER(COALESCE(mv.mee_codigo,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(m.descripcion,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(mv.lote_ref,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(mv.batch_ref,'')) LIKE ? ESCAPE '\\' OR "
            "LOWER(COALESCE(mv.observaciones,'')) LIKE ? ESCAPE '\\')",
        )
        params.extend([like]*5)
    where_sql = ' AND '.join(where)
    conn = get_db(); c = conn.cursor()
    # Total
    try:
        total = c.execute(
            f"""SELECT COUNT(*) FROM movimientos_mee mv
                LEFT JOIN maestro_mee m ON mv.mee_codigo = m.codigo
                WHERE {where_sql}""", params,
        ).fetchone()[0]
    except Exception:
        total = 0
    # Datos
    sql = (
        f"""SELECT mv.id, mv.mee_codigo, COALESCE(m.descripcion,'') as descripcion,
                   mv.tipo, mv.cantidad, mv.unidad, mv.lote_ref, mv.batch_ref,
                   mv.responsable, mv.observaciones, mv.fecha, mv.anulado
            FROM movimientos_mee mv
            LEFT JOIN maestro_mee m ON mv.mee_codigo = m.codigo
            WHERE {where_sql}
            ORDER BY mv.fecha DESC, mv.id DESC
            LIMIT ? OFFSET ?"""
    )
    try:
        c.execute(sql, params + [limit, offset])
        cols = [d[0] for d in c.description]
        rows = [dict(zip(cols, r)) for r in c.fetchall()]
    except Exception:
        rows = []
    return jsonify({
        'movimientos': rows, 'total': total,
        'limit': limit, 'offset': offset,
        'has_more': (offset + len(rows)) < total,
    })


@bp.route('/api/mee/recalcular-stock', methods=['POST'])
def mee_recalcular_stock():
    """Sprint MEE PRO · 20-may-2026 · anti-drift.

    `maestro_mee.stock_actual` es un cache · puede drifear contra
    SUM(movimientos_mee). Este endpoint recalcula el stock real desde
    los movimientos y actualiza la columna.

    Body: {codigo: str | null}
      Si codigo se pasa · solo recalcula ese.
      Si null · recalcula TODOS los MEE activos (admin only).

    Devuelve por código: stock_anterior, stock_calculado, delta.
    """
    u, err, code = _require_session()
    if err:
        return err, code
    body = request.get_json(silent=True) or {}
    codigo_req = (body.get('codigo') or '').strip()
    if not codigo_req and u not in ADMIN_USERS:
        return jsonify({'error': 'recálculo masivo solo admin'}), 403

    conn = get_db(); c = conn.cursor()
    # Calcular stock real desde movimientos_mee (excluye anulados)
    where_mee = "WHERE COALESCE(estado,'Activo')='Activo'"
    params = []
    if codigo_req:
        where_mee += " AND codigo = ?"
        params.append(codigo_req)
    mps = c.execute(
        f"SELECT codigo, stock_actual, unidad FROM maestro_mee {where_mee}",
        params,
    ).fetchall()
    cambios = []
    for r in mps:
        cod, stock_ant, unidad = r
        # SUM de movs (Entrada+, Salida-, Ajuste se trata como delta ya cargado)
        sum_row = c.execute(
            """SELECT COALESCE(SUM(CASE
                   WHEN tipo='Entrada' THEN cantidad
                   WHEN tipo='Salida'  THEN -cantidad
                   WHEN tipo='Ajuste'  THEN cantidad
                   ELSE 0
               END), 0)
               FROM movimientos_mee
               WHERE mee_codigo = ? AND anulado = 0""",
            (cod,),
        ).fetchone()
        stock_calc = max(float(sum_row[0] or 0), 0)
        stock_ant_f = float(stock_ant or 0)
        delta = round(stock_calc - stock_ant_f, 2)
        if abs(delta) >= 0.5:
            c.execute(
                "UPDATE maestro_mee SET stock_actual = ? WHERE codigo = ?",
                (stock_calc, cod),
            )
            cambios.append({
                'codigo': cod, 'stock_anterior': stock_ant_f,
                'stock_calculado': stock_calc, 'delta': delta,
                'unidad': unidad or 'und',
            })
    # audit_log
    try:
        from json import dumps as _jdumps
        c.execute(
            """INSERT INTO audit_log (usuario,accion,tabla,registro_id,detalle,ip,fecha)
               VALUES (?,?,?,?,?,?,datetime('now','-5 hours'))""",
            (u, 'RECALCULAR_STOCK_MEE', 'maestro_mee',
             codigo_req or '_BULK_',
             _jdumps({'cambios': len(cambios), 'codigos': [c['codigo'] for c in cambios][:30]}),
             request.remote_addr or ''),
        )
    except Exception:
        pass
    conn.commit()
    return jsonify({
        'ok': True,
        'evaluados': len(mps),
        'recalculados': len(cambios),
        'cambios': cambios[:100],
        'mensaje': (f'{len(cambios)} MEE actualizado(s) de {len(mps)} evaluados'
                    if cambios else f'Sin drift · {len(mps)} MEE coherentes'),
    })

@bp.route('/api/mee/alertas', methods=['GET'])
def mee_alertas_list():
    """Alertas MEE: bajo minimo + posible obsolescencia (sin mov >90 dias)."""
    from datetime import date, timedelta
    conn = get_db(); c = conn.cursor()
    hace90 = (date.today() - timedelta(days=90)).isoformat()

    c.execute("""SELECT m.codigo, m.descripcion, m.categoria, m.stock_actual, m.stock_minimo, m.unidad
                 FROM maestro_mee m
                 WHERE m.estado='Activo' AND m.stock_minimo>0 AND m.stock_actual < m.stock_minimo
                 ORDER BY (m.stock_actual / m.stock_minimo) ASC""")
    bajo_minimo = [{'codigo': r[0], 'descripcion': r[1], 'categoria': r[2],
                    'stock_actual': r[3], 'stock_minimo': r[4], 'unidad': r[5],
                    'ratio': round(r[3]/r[4], 2) if r[4] else 0} for r in c.fetchall()]

    c.execute("""SELECT m.codigo, m.descripcion, m.categoria, m.stock_actual, m.unidad,
                        MAX(mv.fecha) as ultimo_mov
                 FROM maestro_mee m
                 LEFT JOIN movimientos_mee mv ON m.codigo=mv.mee_codigo AND mv.anulado=0
                 WHERE m.estado='Activo' AND m.stock_actual>0
                 GROUP BY m.codigo
                 HAVING ultimo_mov IS NULL OR ultimo_mov < ?
                 ORDER BY ultimo_mov ASC LIMIT 15""", (hace90,))
    obsolescencia = [{'codigo': r[0], 'descripcion': r[1], 'categoria': r[2],
                      'stock_actual': r[3], 'unidad': r[4], 'ultimo_mov': r[5] or 'Nunca'}
                     for r in c.fetchall()]

    c.execute("SELECT COUNT(*) FROM maestro_mee WHERE estado='Activo'")
    total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM maestro_mee WHERE estado='Activo' AND stock_minimo>0 AND stock_actual<stock_minimo")
    n_bajo = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM movimientos_mee WHERE anulado=0 AND fecha>=date('now', '-5 hours', '-7 days')")
    mov_sem = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM movimientos_mee WHERE tipo='Entrada' AND anulado=0 AND fecha>=date('now', '-5 hours', '-30 days')")
    ent_mes = c.fetchone()[0]
    return jsonify({
        'bajo_minimo': bajo_minimo, 'obsolescencia': obsolescencia,
        'resumen': {'total_mee': total, 'bajo_minimo': n_bajo,
                    'movimientos_semana': mov_sem, 'entradas_mes': ent_mes}
    })

@bp.route('/api/mee/trazabilidad', methods=['GET'])
def mee_trazabilidad():
    """Trazabilidad MEE: batch->empaque consumido  |  codigo->historial de batches."""
    batch  = request.args.get('batch','').strip()
    codigo = request.args.get('codigo','').strip()
    conn = get_db(); c = conn.cursor()

    if batch:
        c.execute("""SELECT mv.id, mv.mee_codigo, COALESCE(m.descripcion,'') as descripcion,
                            m.categoria, mv.cantidad, mv.unidad, mv.responsable, mv.fecha, mv.observaciones
                     FROM movimientos_mee mv LEFT JOIN maestro_mee m ON mv.mee_codigo=m.codigo
                     WHERE mv.batch_ref LIKE ? AND mv.tipo='Salida' AND mv.anulado=0
                     ORDER BY mv.fecha""", (f'%{batch}%',))
        cols = [d[0] for d in c.description]
        consumos = [dict(zip(cols, r)) for r in c.fetchall()]
        return jsonify({'tipo': 'batch', 'referencia': batch, 'consumos': consumos, 'total': len(consumos)})

    elif codigo:
        c.execute("""SELECT mv.id, mv.tipo, mv.batch_ref, mv.lote_ref, mv.cantidad, mv.unidad,
                            mv.responsable, mv.fecha, mv.observaciones
                     FROM movimientos_mee mv
                     WHERE mv.mee_codigo=? AND mv.anulado=0
                     ORDER BY mv.fecha DESC LIMIT 100""", (codigo,))
        cols = [d[0] for d in c.description]
        historial = [dict(zip(cols, r)) for r in c.fetchall()]
        return jsonify({'tipo': 'mee', 'referencia': codigo, 'historial': historial, 'total': len(historial)})

    return jsonify({'error': 'Proporcione parametro batch o codigo'}), 400

@bp.route('/api/mee/anular/<int:mov_id>', methods=['POST'])
def mee_anular_movimiento(mov_id):
    """Anula un movimiento MEE revirtiendo el impacto en stock · auditado."""
    user, err, code = _require_planta_write()
    if err:
        return err, code
    payload = request.json or {}
    motivo = payload.get('motivo','').strip()
    if len(motivo) < 5:
        return jsonify({'error': 'Motivo (≥5 chars) obligatorio'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM movimientos_mee WHERE id=?", (mov_id,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'No encontrado'}), 404
    cols_mv = [d[0] for d in c.description]
    mv = dict(zip(cols_mv, row))
    if mv['anulado']:
        return jsonify({'error': 'Ya anulado'}), 400
    # Audit zero-error 2-may-2026: tipo='Ajuste' (legacy pre-may-2026) NO se
    # puede anular automáticamente porque el signo del ajuste se perdía. Para
    # revertir un ajuste legacy, hay que crear un nuevo ajuste manual con la
    # operación inversa.
    if mv['tipo'] == 'Ajuste':
        return jsonify({
            'error': 'No se puede anular un ajuste legacy (signo no preservado). '
                      'Para revertir, crea un nuevo ajuste manual con la operación inversa.',
            'codigo': 'AJUSTE_LEGACY_NO_ANULABLE',
        }), 422
    if mv['tipo'] == 'Entrada':
        c.execute("UPDATE maestro_mee SET stock_actual=MAX(0,stock_actual-?) WHERE codigo=?",
                  (mv['cantidad'], mv['mee_codigo']))
    elif mv['tipo'] == 'Salida':
        c.execute("UPDATE maestro_mee SET stock_actual=stock_actual+? WHERE codigo=?",
                  (mv['cantidad'], mv['mee_codigo']))
    else:
        return jsonify({'error': f"tipo desconocido '{mv['tipo']}' · no puede anular"}), 400
    c.execute("UPDATE movimientos_mee SET anulado=1, observaciones=observaciones||? WHERE id=?",
              (f' [ANULADO por {user}: {motivo}]', mov_id))
    audit_log(c, usuario=user, accion='ANULAR_MOV_MEE', tabla='movimientos_mee',
              registro_id=mov_id,
              antes={'tipo': mv['tipo'], 'cantidad': mv['cantidad'],
                     'mee_codigo': mv['mee_codigo'], 'anulado': 0},
              despues={'anulado': 1, 'motivo': motivo},
              detalle=f"Anuló movimiento MEE #{mov_id} ({mv['tipo']} {mv['cantidad']} de {mv['mee_codigo']}) · motivo: {motivo}")
    conn.commit()
    return jsonify({'ok': True, 'message': f'Movimiento #{mov_id} anulado y stock revertido'})


# ════════════════════════════════════════════════════════════════════════
# MEE ampliado (29-abr-2026): proveedor, ajustar, historico, eliminar,
# bulk import desde Excel — paridad de funcionalidades con MP.
# ════════════════════════════════════════════════════════════════════════

@bp.route('/api/mee/<codigo>', methods=['GET', 'PUT', 'DELETE'])
def mee_item_detalle(codigo):
    """GET: detalle de un MEE con stock, ultimo mov.
    PUT: actualiza descripcion / categoria / unidad / proveedor.
    DELETE: archiva (estado='Archivado'), no borra fisicamente.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user','')
    conn = get_db(); c = conn.cursor()
    if request.method == 'DELETE':
        # P1 audit 26-may · audit_log archivado (soft delete) · MEE master
        # snapshot antes del UPDATE
        _snap_row = c.execute(
            "SELECT codigo, descripcion, categoria, proveedor, estado FROM maestro_mee WHERE codigo=?",
            (codigo,)).fetchone()
        c.execute("UPDATE maestro_mee SET estado='Archivado' WHERE codigo=?", (codigo,))
        if c.rowcount == 0:
            return jsonify({'error':'No encontrado'}), 404
        if _snap_row:
            try:
                audit_log(c, usuario=user, accion='ARCHIVAR_MEE', tabla='maestro_mee',
                          registro_id=codigo,
                          antes={'estado': _snap_row[4], 'descripcion': _snap_row[1] or '',
                                 'categoria': _snap_row[2] or '', 'proveedor': _snap_row[3] or ''},
                          despues={'estado': 'Archivado'},
                          detalle=f'MEE {codigo} archivado')
            except Exception as _ae:
                import logging as _lg
                _lg.getLogger('inventario').warning('audit ARCHIVAR_MEE fallo: %s', _ae)
        conn.commit()
        return jsonify({'ok': True, 'archivado': codigo})
    if request.method == 'PUT':
        d = request.json or {}
        sets, params = [], []
        for f in ('descripcion','categoria','unidad','proveedor','fabricante','stock_minimo'):
            if f in d:
                sets.append(f'{f}=?'); params.append(d[f])
        if not sets:
            return jsonify({'error':'Nada que actualizar'}), 400
        params.append(codigo)
        c.execute(f"UPDATE maestro_mee SET {', '.join(sets)} WHERE codigo=?", params)
        if c.rowcount == 0:
            return jsonify({'error':'No encontrado'}), 404
        conn.commit()
        return jsonify({'ok': True, 'codigo': codigo})
    # GET
    row = c.execute("""
        SELECT codigo, descripcion, categoria, proveedor, fabricante, estado,
               stock_actual, stock_minimo, unidad, fecha_creacion
        FROM maestro_mee WHERE codigo=?
    """, (codigo,)).fetchone()
    if not row:
        return jsonify({'error':'No encontrado'}), 404
    cols = [d[0] for d in c.description]
    return jsonify(dict(zip(cols, row)))


@bp.route('/api/mee/<codigo>/proveedor', methods=['PUT'])
def mee_set_proveedor(codigo):
    """Cambiar proveedor de un MEE (igual que MP)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    proveedor = (d.get('proveedor') or '').strip()
    conn = get_db(); c = conn.cursor()
    # P1 audit 26-may · snapshot proveedor anterior para audit_log
    prov_ant_row = c.execute("SELECT proveedor FROM maestro_mee WHERE codigo=?", (codigo,)).fetchone()
    prov_ant = prov_ant_row[0] if prov_ant_row else None
    c.execute("UPDATE maestro_mee SET proveedor=? WHERE codigo=?", (proveedor, codigo))
    if c.rowcount == 0:
        return jsonify({'error':'No encontrado'}), 404
    try:
        audit_log(c, usuario=user, accion='UPDATE_MEE_PROVEEDOR', tabla='maestro_mee',
                  registro_id=codigo,
                  antes={'proveedor': prov_ant or ''},
                  despues={'proveedor': proveedor},
                  detalle=f'MEE {codigo} proveedor: {prov_ant!r} → {proveedor!r}')
    except Exception as _ae:
        import logging as _lg
        _lg.getLogger('inventario').warning('audit UPDATE_MEE_PROVEEDOR fallo: %s', _ae)
    conn.commit()
    return jsonify({'ok': True, 'codigo': codigo, 'proveedor': proveedor})


@bp.route('/api/mee/<codigo>/stock-minimo', methods=['PUT'])
def mee_set_stock_minimo(codigo):
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user', '')
    d = request.json or {}
    # P1 audit 26-may · validate_money + audit_log · antes float() pelado
    # aceptaba NaN/Inf/negativos · stock_minimo es threshold de alerta.
    from http_helpers import validate_money as _vm_sm
    nuevo, _err = _vm_sm(d.get('stock_minimo', 0), allow_zero=True,
                          max_value=10_000_000, field_name='stock_minimo')
    if _err:
        return jsonify(_err), 400
    conn = get_db(); c = conn.cursor()
    ant_row = c.execute("SELECT stock_minimo FROM maestro_mee WHERE codigo=?", (codigo,)).fetchone()
    ant = ant_row[0] if ant_row else None
    c.execute("UPDATE maestro_mee SET stock_minimo=? WHERE codigo=?", (nuevo, codigo))
    if c.rowcount == 0:
        return jsonify({'error':'No encontrado'}), 404
    try:
        audit_log(c, usuario=user, accion='UPDATE_MEE_STOCK_MIN', tabla='maestro_mee',
                  registro_id=codigo,
                  antes={'stock_minimo': ant},
                  despues={'stock_minimo': nuevo},
                  detalle=f'MEE {codigo} stock_minimo: {ant} → {nuevo}')
    except Exception as _ae:
        import logging as _lg
        _lg.getLogger('inventario').warning('audit UPDATE_MEE_STOCK_MIN fallo: %s', _ae)
    conn.commit()
    return jsonify({'ok': True, 'codigo': codigo, 'stock_minimo': nuevo})


@bp.route('/api/mee/<codigo>/ajustar', methods=['POST'])
def mee_ajustar_stock(codigo):
    """Ajusta stock manual con motivo (auditado en movimientos_mee)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user','')
    d = request.json or {}
    # FIX P1 audit 24-may-2026 · validate_money para rechazar NaN/Inf/negativo.
    # allow_zero=True permite vaciar stock a 0.
    from http_helpers import validate_money
    cantidad_nueva, err = validate_money(d.get('cantidad_nueva', 0),
                                          allow_zero=True, max_value=10_000_000,
                                          field_name='cantidad_nueva')
    if err:
        return jsonify(err), 400
    motivo = (d.get('motivo') or '').strip()
    if not motivo:
        return jsonify({'error': 'motivo requerido para ajuste'}), 400
    conn = get_db(); c = conn.cursor()
    row = c.execute("SELECT stock_actual FROM maestro_mee WHERE codigo=?", (codigo,)).fetchone()
    if not row:
        return jsonify({'error':'No encontrado'}), 404
    # FIX · 21-may-2026 · stock_anterior REAL desde SUM(movimientos_mee)
    # No usar cache (puede estar drifteado) · ajuste compounding sino.
    try:
        real_row = c.execute(
            """SELECT COALESCE(SUM(CASE
                   WHEN tipo='Entrada' THEN cantidad
                   WHEN tipo='Salida'  THEN -cantidad
                   WHEN tipo='Ajuste'  THEN cantidad
                   ELSE 0 END), 0)
               FROM movimientos_mee WHERE mee_codigo=? AND COALESCE(anulado,0)=0""",
            (codigo,),
        ).fetchone()
        stock_anterior = max(float(real_row[0] or 0), 0)
    except Exception:
        stock_anterior = float(row[0] or 0)
    delta = cantidad_nueva - stock_anterior
    c.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?", (cantidad_nueva, codigo))
    # Audit zero-error 2-may-2026: usar Entrada/Salida según signo del delta
    # (no 'Ajuste' que perdía la dirección · provocaba drift permanente entre
    # stock_actual y SUM(movimientos_mee)). Observaciones marca AJUSTE MANUAL
    # para que reportes financieros lo distingan de movimientos operativos.
    if delta == 0:
        # Sin cambio · no insertar movimiento (idempotente)
        pass
    else:
        tipo_mov = 'Entrada' if delta > 0 else 'Salida'
        obs_msg = (f'AJUSTE MANUAL: {stock_anterior} → {cantidad_nueva} '
                    f'({"+" if delta>=0 else ""}{delta}). {motivo}')
        c.execute("""
            INSERT INTO movimientos_mee (mee_codigo, tipo, cantidad, responsable, observaciones)
            VALUES (?, ?, ?, ?, ?)
        """, (codigo, tipo_mov, abs(delta), user, obs_msg))
    audit_log(c, usuario=user, accion='AJUSTAR_STOCK_MEE', tabla='maestro_mee',
              registro_id=codigo,
              antes={'stock_actual': stock_anterior},
              despues={'stock_actual': cantidad_nueva, 'delta': delta, 'motivo': motivo},
              detalle=f"Ajustó stock MEE {codigo}: {stock_anterior} → {cantidad_nueva} (Δ {delta:+}) · {motivo}")
    conn.commit()
    return jsonify({'ok': True, 'codigo': codigo, 'stock_anterior': stock_anterior,
                    'stock_nuevo': cantidad_nueva, 'delta': delta})


@bp.route('/api/mee/<codigo>/historico', methods=['GET'])
def mee_historico_item(codigo):
    """Histórico de movimientos de UN item MEE (ordenado descendente)."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT id, tipo, cantidad, unidad, lote_ref, batch_ref,
               responsable, observaciones, fecha, COALESCE(anulado,0) as anulado
        FROM movimientos_mee
        WHERE mee_codigo=?
        ORDER BY fecha DESC, id DESC
        LIMIT 200
    """, (codigo,)).fetchall()
    cols = [d[0] for d in c.description]
    return jsonify({'codigo': codigo, 'movimientos': [dict(zip(cols, r)) for r in rows]})


@bp.route('/api/mee/import-bulk', methods=['POST'])
def mee_import_bulk():
    """Importa o actualiza MEE en lote desde JSON.

    Body: {items: [{codigo, descripcion, categoria, presentacion, stock,
                    proveedor?, unidad?, stock_minimo?}], modo: 'upsert'|'replace'}

    upsert (default): si codigo existe, UPDATE descripcion/categoria/stock;
                       si no existe, INSERT. No borra los que no aparezcan.
    replace: archiva (estado='Archivado') los items NO incluidos.
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    user = session.get('compras_user','')
    d = request.json or {}
    items = d.get('items') or []
    modo = (d.get('modo') or 'upsert').strip()
    if not items:
        return jsonify({'error': 'items vacio'}), 400
    conn = get_db(); c = conn.cursor()
    insertados = 0; actualizados = 0
    codigos_recibidos = set()
    for it in items:
        codigo = (it.get('codigo') or '').strip()
        if not codigo: continue
        codigos_recibidos.add(codigo)
        descripcion = (it.get('descripcion') or '').strip()
        categoria = (it.get('categoria') or 'Otro').strip()
        unidad = (it.get('unidad') or 'und').strip()
        proveedor = (it.get('proveedor') or '').strip()
        # FIX P1 audit 24-may-2026 · validar stock numérico, rechazar
        # NaN/Inf/abs. En bulk no abortamos el batch entero por un row
        # malo · saltamos al siguiente y reportamos en respuesta.
        from http_helpers import validate_money
        stock, err_s = validate_money(it.get('stock', 0),
                                       allow_zero=True, max_value=10_000_000)
        if err_s:
            stock = 0
        stock_min, err_sm = validate_money(it.get('stock_minimo', 1000),
                                            allow_zero=True, max_value=10_000_000)
        if err_sm:
            stock_min = 1000
        existing = c.execute("SELECT codigo, stock_actual FROM maestro_mee WHERE codigo=?", (codigo,)).fetchone()
        if existing:
            stock_anterior = float(existing[1] or 0)
            c.execute("""
                UPDATE maestro_mee SET
                  descripcion=?, categoria=?, unidad=?, proveedor=?,
                  stock_actual=?, stock_minimo=?, estado='Activo'
                WHERE codigo=?
            """, (descripcion, categoria, unidad, proveedor, stock, stock_min, codigo))
            # FIX-B4 13-may-2026: antes insertaba tipo='Ajuste' con
            # abs(delta) · pero stock_mee_calculated trata 'Ajuste' como
            # +cantidad (positivo). Si stock_nuevo < stock_anterior, el
            # UPDATE bajaba el stock_actual pero el INSERT subía el kardex
            # calculado · drift permanente de 2×|delta|. Ahora usamos
            # Entrada/Salida según signo del delta (igual patrón que
            # mee_ajustar_stock línea 4910).
            if abs(stock - stock_anterior) > 0.01:
                delta = stock - stock_anterior
                tipo_mov = 'Entrada' if delta > 0 else 'Salida'
                c.execute("""
                    INSERT INTO movimientos_mee (mee_codigo, tipo, cantidad, responsable, observaciones)
                    VALUES (?, ?, ?, ?, ?)
                """, (codigo, tipo_mov, abs(delta), user,
                      f'[Import Excel] {stock_anterior} → {stock} ({tipo_mov.lower()})'))
            actualizados += 1
        else:
            c.execute("""
                INSERT INTO maestro_mee (codigo, descripcion, categoria, unidad, proveedor,
                                         stock_actual, stock_minimo, estado, fecha_creacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'Activo', datetime('now', '-5 hours'))
            """, (codigo, descripcion, categoria, unidad, proveedor, stock, stock_min))
            c.execute("""
                INSERT INTO movimientos_mee (mee_codigo, tipo, cantidad, responsable, observaciones)
                VALUES (?, 'Entrada', ?, ?, ?)
            """, (codigo, stock, user, '[Import Excel] inventario inicial'))
            insertados += 1
    archivados = 0
    if modo == 'replace':
        # Archivar los que NO vinieron (estado='Activo' no recibidos)
        rows = c.execute("SELECT codigo FROM maestro_mee WHERE COALESCE(estado,'Activo')='Activo'").fetchall()
        for (cod,) in rows:
            if cod not in codigos_recibidos:
                c.execute("UPDATE maestro_mee SET estado='Archivado' WHERE codigo=?", (cod,))
                archivados += 1
    # INVIMA-FIX · 21-may-2026 · audit_log obligatorio (BPM Resolución 2214/2021)
    try:
        audit_log(c, usuario=user, accion='IMPORT_BULK_MEE',
                  tabla='maestro_mee', registro_id='_BULK_',
                  despues={'insertados': insertados, 'actualizados': actualizados,
                           'archivados': archivados, 'modo': modo,
                           'items_count': len(items)})
    except Exception as _e:
        __import__('logging').getLogger('inventario').warning('audit IMPORT_BULK_MEE: %s', _e)
    conn.commit()
    return jsonify({
        'ok': True,
        'insertados': insertados,
        'actualizados': actualizados,
        'archivados': archivados,
        'total_recibidos': len(items),
    })


@bp.route('/api/mee/categorias', methods=['GET'])
def mee_categorias():
    """Lista categorias distintas del catalogo MEE con conteo."""
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT COALESCE(categoria,'Sin categoria') as cat,
               COUNT(*) as n,
               COALESCE(SUM(stock_actual),0) as stock_total
        FROM maestro_mee
        WHERE COALESCE(estado,'Activo')='Activo'
        GROUP BY cat
        ORDER BY cat
    """).fetchall()
    cats = [{'categoria': r[0], 'count': r[1], 'stock_total': float(r[2] or 0)} for r in rows]
    return jsonify({'categorias': cats})


# ═══════════════════════════════════════════════════════════════
#  ACONDICIONAMIENTO + LIBERACIÓN — Fase 4
# ═══════════════════════════════════════════════════════════════

def _init_acondicionamiento():
    conn = get_db(); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS acondicionamiento (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        produccion_id       INTEGER DEFAULT 0,
        envasado_id         INTEGER DEFAULT 0,
        lote                TEXT NOT NULL DEFAULT '',
        producto            TEXT NOT NULL DEFAULT '',
        cantidad_batch_g    REAL DEFAULT 0,
        unidades_producidas INTEGER DEFAULT 0,
        presentacion        TEXT DEFAULT '',
        mee_consumido       TEXT DEFAULT '[]',
        fecha               TEXT DEFAULT (date('now', '-5 hours')),
        operador            TEXT DEFAULT '',
        observaciones       TEXT DEFAULT '',
        estado              TEXT DEFAULT 'En proceso',
        creado_en           DATETIME DEFAULT (datetime('now', '-5 hours'))
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS liberaciones (
        id                      INTEGER PRIMARY KEY AUTOINCREMENT,
        acondicionamiento_id    INTEGER DEFAULT 0,
        lote                    TEXT NOT NULL DEFAULT '',
        producto                TEXT NOT NULL DEFAULT '',
        unidades                INTEGER DEFAULT 0,
        presentacion            TEXT DEFAULT '',
        fecha_produccion        TEXT DEFAULT '',
        fecha_liberacion        TEXT DEFAULT '',
        aprobado_por            TEXT DEFAULT '',
        cliente                 TEXT DEFAULT '',
        destino                 TEXT DEFAULT 'ANIMUS',
        observaciones           TEXT DEFAULT '',
        estado                  TEXT DEFAULT 'Pendiente CC',
        creado_en               DATETIME DEFAULT (datetime('now', '-5 hours'))
    )""")
    # Nuevas columnas — múltiples presentaciones y flujo liberación→stock_pt.
    # Usamos safe_alter (database.py) que distingue "columna ya existe"
    # (benigno) de errores reales (loguea + relanza).
    from database import safe_alter
    for _sql in [
        "ALTER TABLE acondicionamiento ADD COLUMN sku TEXT DEFAULT ''",
        "ALTER TABLE acondicionamiento ADD COLUMN precio_base REAL DEFAULT 0",
        "ALTER TABLE liberaciones ADD COLUMN sku TEXT DEFAULT ''",
        "ALTER TABLE liberaciones ADD COLUMN precio_base REAL DEFAULT 0",
    ]:
        safe_alter(conn, _sql)
    conn.commit()

_init_acondicionamiento()

# ══════════════════════════════════════════════════════════════════
# ENVASADO — Paso entre Produccion y Acondicionamiento
# ══════════════════════════════════════════════════════════════════

@bp.route('/api/producciones/sin-envasar', methods=['GET'])
def producciones_sin_envasar():
    """Cola de producciones sin registro de envasado vinculado."""
    conn = get_db(); c = conn.cursor()
    c.execute("""
        SELECT p.id, p.lote, p.producto, p.cantidad, p.fecha, p.operador, p.presentacion
        FROM producciones p
        LEFT JOIN envasado e ON e.produccion_id = p.id
        WHERE e.id IS NULL
          AND COALESCE(p.estado,'') NOT IN ('cancelado','Cancelado')
        ORDER BY p.id DESC LIMIT 100
    """)
    cols = ['id','lote','producto','cantidad_kg','fecha','operador','presentacion']
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({'cola': rows})

@bp.route('/api/envasado', methods=['GET', 'POST'])
def envasado_list():
    # P0 audit 26-may-2026 · sin gate · POST creaba envasado + UPDATE MEE stock.
    _u_sess, _err_s, _code_s = _require_session()
    if _err_s:
        return _err_s, _code_s
    if request.method == 'POST':
        _u_w, _err_w, _code_w = _require_planta_write()
        if _err_w:
            return _err_w, _code_w
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.get_json(silent=True) or {}
        produccion_id = int(d.get('produccion_id') or 0)
        lote = d.get('lote', '').strip()
        producto = d.get('producto', '').strip()
        presentacion = d.get('presentacion', '').strip()
        batch_g = float(d.get('batch_g', 0) or 0)
        unidades = int(d.get('unidades', 0) or 0)
        envase_codigo = d.get('envase_codigo', '').strip()
        tapa_codigo = d.get('tapa_codigo', '').strip()
        operador = d.get('operador', session.get('compras_user', '')).strip()
        fecha = d.get('fecha', datetime.now().strftime('%Y-%m-%d'))
        obs = d.get('observaciones', '').strip()
        area_codigo = (d.get('area_codigo') or '').strip()

        if not lote or not producto:
            return jsonify({'error': 'lote y producto son requeridos'}), 400
        if unidades <= 0:
            return jsonify({'error': 'unidades debe ser > 0'}), 400
        # Semi-auto envasado (9-jun) · GATE de área limpia (avisar+override · M5): el
        # área de envasado asignada debe estar LIMPIA (areas_planta.estado='libre')
        # antes de envasar. Pre-llenada en el modal vía /api/planta/envasado/sugerencias.
        if area_codigo:
            _ar = c.execute(
                "SELECT COALESCE(estado,'') FROM areas_planta WHERE codigo=? AND COALESCE(activo,1)=1",
                (area_codigo,)).fetchone()
            if _ar and _ar[0] != 'libre' and not bool(d.get('override_area')):
                return jsonify({
                    'warning': f'El área {area_codigo} NO está limpia (estado: {_ar[0] or "?"}). '
                               f'Limpiá el área (rótulo F02) o confirmá para envasar igual.',
                    'requiere_override': True, 'bloqueo': 'area_no_limpia',
                }), 409

        # ─── PRE-CHECK MEE ──────────────────────────────────────────────────
        # Validar que los códigos existan en maestro_mee (anti-typo) y que
        # haya stock suficiente. Si falta cualquiera → 422 sin escribir nada.
        # Antes: si el código no existía, se hacía continue silencioso y el
        # INSERT en envasado quedaba sin descuento, dejando data inconsistente.
        plan_mee = []  # [(codigo, cant, descripcion, stock_actual, stock_minimo)]
        errores_mee = []
        for tipo_mee, codigo_mee, cant in [
            ('envase', envase_codigo, unidades),
            ('tapa', tapa_codigo, unidades),
        ]:
            if not codigo_mee:
                continue  # opcional: producto sin tapa, etc.
            row = c.execute(
                "SELECT stock_actual, stock_minimo, descripcion, estado "
                "FROM maestro_mee WHERE codigo=?",
                (codigo_mee,)
            ).fetchone()
            if not row:
                errores_mee.append({
                    'tipo': tipo_mee,
                    'codigo': codigo_mee,
                    'error': 'código no existe en maestro_mee',
                })
                continue
            stock_actual, stock_minimo, descripcion, estado = row[0], row[1], row[2], row[3]
            if estado != 'Activo':
                errores_mee.append({
                    'tipo': tipo_mee,
                    'codigo': codigo_mee,
                    'error': f'item está {estado}, no Activo',
                })
                continue
            if (stock_actual or 0) < cant:
                errores_mee.append({
                    'tipo': tipo_mee,
                    'codigo': codigo_mee,
                    'descripcion': descripcion,
                    'stock_disponible': stock_actual,
                    'requerido': cant,
                    'falta': cant - (stock_actual or 0),
                    'error': 'stock insuficiente',
                })
                continue
            plan_mee.append((codigo_mee, cant, descripcion or '',
                             stock_actual, stock_minimo))

        if errores_mee:
            return jsonify({
                'error': 'No se puede registrar el envasado',
                'detalle': 'Códigos MEE inválidos o sin stock suficiente',
                'errores': errores_mee,
                'mensaje': (
                    'Verifica los códigos en el dropdown y que haya stock. '
                    'Si necesitás más MEE, crea OC en /compras.'
                ),
            }), 422

        # ─── ESCRITURA TRANSACCIONAL ─────────────────────────────────────────
        try:
            c.execute("""INSERT INTO envasado
                (produccion_id, lote, producto, presentacion, batch_g, unidades,
                 envase_codigo, tapa_codigo, operador, fecha, estado, observaciones, area_codigo)
                VALUES (?,?,?,?,?,?,?,?,?,?,'Completado',?,?)""",
                (produccion_id, lote, producto, presentacion, batch_g, unidades,
                 envase_codigo, tapa_codigo, operador, fecha, obs, area_codigo))
            nuevo_id = c.lastrowid
            # Semi-auto · marcar el área asignada como ocupada (deja de estar libre).
            if area_codigo:
                c.execute("UPDATE areas_planta SET estado='ocupada' "
                          "WHERE codigo=? AND COALESCE(activo,1)=1 AND estado='libre'",
                          (area_codigo,))

            alertas_mee = []
            for codigo_mee, cant, descripcion, stock_actual, stock_minimo in plan_mee:
                nuevo_stock = max(0, (stock_actual or 0) - cant)
                c.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?",
                          (nuevo_stock, codigo_mee))
                c.execute("""INSERT INTO movimientos_mee
                    (mee_codigo, tipo, cantidad, lote_ref, batch_ref, observaciones, responsable, fecha)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (codigo_mee, 'Salida', cant,
                     str(produccion_id) if produccion_id else lote,
                     lote,
                     'Envasado ' + lote + ' - ' + producto + ' ' + presentacion,
                     operador, fecha))
                # Sebastian (1-may-2026): marcar consumido_at en checklist para
                # evitar doble descuento al completar producción.
                if produccion_id:
                    try:
                        c.execute("""
                            UPDATE produccion_checklist
                               SET consumido_at = ?,
                                   consumido_por = ?,
                                   cantidad_consumida_real = ?,
                                   consumido_contexto = 'envasado',
                                   actualizado_at = datetime('now', '-5 hours')
                             WHERE produccion_id = ?
                               AND mee_codigo_asignado = ?
                               AND COALESCE(consumido_at,'') = ''
                        """, (datetime.now().isoformat(timespec='seconds'),
                              operador, cant, produccion_id, codigo_mee))
                    except Exception:
                        pass  # tabla puede no tener las columnas (DB legacy)
                if (stock_minimo or 0) > 0 and nuevo_stock < stock_minimo:
                    deficit = stock_minimo - nuevo_stock
                    alertas_mee.append({
                        'codigo': codigo_mee, 'nombre': descripcion,
                        'stock': nuevo_stock, 'minimo': stock_minimo,
                        'deficit': deficit,
                    })

            # Reemplazo MyBatch · 9-jun-2026 · LEGAJO OF AUTOMÁTICO de envasado: al
            # envasar nace el EBR de fase ENVASADO (la "Orden de Envasado") si el
            # producto tiene MBR aprobado con pasos de envasado. Auto-gateado:
            # NO_MBR_APROBADO → no-op (no bloquea · idéntico a antes para productos sin
            # MBR). Idempotente por (produccion_id, lote): N presentaciones del mismo
            # lote = 1 solo legajo. Espeja el hook de fabricación (línea ~2354).
            try:
                from blueprints.brd import crear_ebr_desde_mbr
                _re = crear_ebr_desde_mbr(
                    c, producto_nombre=producto, lote=lote,
                    produccion_id=(produccion_id or None), usuario=operador,
                    fase='envasado', area_codigo=area_codigo)
                if _re.get('ok') and not _re.get('reusado'):
                    try:
                        from database import audit_log as _ale
                        _ale(c, usuario=operador or 'sistema', accion='CREAR_EBR_OF_AUTO',
                             tabla='ebr_ejecuciones', registro_id=str(_re.get('id')),
                             despues={'producto': producto, 'lote': lote,
                                      'fase': 'envasado', 'numero_op': _re.get('numero_op')})
                    except Exception:
                        pass
            except Exception as _eo:
                __import__('logging').getLogger('inventario').warning(
                    'crear EBR OF (envasado) auto fallo (no bloquea envasado): %s', _eo)
        except Exception as _e:
            conn.rollback()
            __import__('logging').getLogger('inventario').error(
                "Envasado FALLÓ tras pre-check OK (rollback): lote=%s err=%s",
                lote, _e, exc_info=True
            )
            return jsonify({
                'error': 'Falla transaccional al registrar envasado',
                'detalle': str(_e),
                'rollback': 'aplicado — no se descontó ningún MEE',
            }), 500

        # Si hay MEE bajo minimo, crear solicitud de compra automatica en Compras
        if alertas_mee:
            try:
                c.execute("SELECT COALESCE(MAX(CAST(SUBSTR(numero,10) AS INTEGER)),0) FROM solicitudes_compra WHERE numero LIKE ?",
                          (f"SOL-{datetime.now().strftime('%Y')}-%",))
                n_sol = (c.fetchone()[0] or 0) + 1
                num_sol = f"SOL-{datetime.now().strftime('%Y')}-{n_sol:04d}"
                obs_sol = 'Alerta automatica envasado ' + lote + ': MEE bajo minimo'
                c.execute("""INSERT INTO solicitudes_compra
                    (numero, fecha, estado, solicitante, urgencia, observaciones,
                     area, empresa, categoria, tipo)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (num_sol, datetime.now().isoformat(), 'Pendiente',
                     operador, 'Alta', obs_sol,
                     'Produccion', 'Espagiria', 'Material de Empaque', 'Compra'))
                for a in alertas_mee:
                    c.execute("""INSERT INTO solicitudes_compra_items
                        (numero, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion)
                        VALUES (?,?,?,?,?,?)""",
                        (num_sol, a['codigo'], a['nombre'], a['deficit'], 'und',
                         'Deficit MEE post-envasado ' + lote))
            except Exception as _e:
                print(f'[envasado] solicitud auto error: {_e}')

        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id, 'alertas_mee': alertas_mee}), 201

    # GET · Sprint Envasado PRO 20-may-2026: paginación + búsqueda
    prod_id = request.args.get('produccion_id', '')
    if prod_id:
        c.execute("SELECT * FROM envasado WHERE produccion_id=? ORDER BY id DESC", (prod_id,))
        cols = [d[0] for d in c.description]
        rows = [dict(zip(cols, r)) for r in c.fetchall()]
        return jsonify({'envasados': rows, 'total': len(rows)})
    try:
        limit = max(1, min(int(request.args.get('limit', 50)), 500))
    except (ValueError, TypeError):
        limit = 50
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    q = (request.args.get('q') or '').strip()
    desde = (request.args.get('desde') or '').strip()
    hasta = (request.args.get('hasta') or '').strip()
    where = ['1=1']
    params = []
    if q:
        qesc = q.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
        where.append("(LOWER(COALESCE(producto,'')) LIKE LOWER(?) ESCAPE '\\' OR LOWER(COALESCE(lote,'')) LIKE LOWER(?) ESCAPE '\\' OR LOWER(COALESCE(operador,'')) LIKE LOWER(?) ESCAPE '\\')")
        params += [f'%{qesc}%', f'%{qesc}%', f'%{qesc}%']
    if desde:
        where.append("fecha >= ?"); params.append(desde)
    if hasta:
        where.append("fecha <= ?"); params.append(hasta + ' 23:59:59')
    where_sql = ' AND '.join(where)
    try:
        total = int(c.execute(f"SELECT COUNT(*) FROM envasado WHERE {where_sql}", params).fetchone()[0] or 0)
    except Exception:
        total = 0
    c.execute(f"SELECT * FROM envasado WHERE {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?", params + [limit, offset])
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({
        'envasados': rows,
        'total': total, 'limit': limit, 'offset': offset,
        'q': q, 'desde': desde, 'hasta': hasta,
    })


@bp.route('/api/envasado/<int:eid>/detalle', methods=['GET'])
def envasado_detalle(eid):
    """Sprint Envasado PRO 20-may-2026 · detalle completo de un envasado:
    header + MEE descontados + costo + producción origen."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM envasado WHERE id=?", (eid,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'envasado no existe'}), 404
    cols = [d[0] for d in c.description]
    env = dict(zip(cols, row))
    # MEE descontados · buscar movimientos_mee con lote igual al envasado
    mee_movs = []
    try:
        mee_rows = c.execute(
            """SELECT material_id, COALESCE(material_nombre,''),
                      cantidad, COALESCE(observaciones,'')
               FROM movimientos_mee
               WHERE tipo='Salida'
                 AND observaciones LIKE ?
               ORDER BY id""",
            (f"%Envasado #{eid}%",),
        ).fetchall()
        mee_movs = [{
            'codigo': r[0], 'descripcion': r[1],
            'unidades': int(r[2] or 0),
            'observaciones': r[3],
        } for r in mee_rows]
    except Exception:
        mee_movs = []
    # Costo estimado MEE
    costo_total = 0.0
    for m in mee_movs:
        try:
            pr_row = c.execute(
                "SELECT COALESCE(precio_unitario, 0) FROM maestro_mee WHERE codigo=?",
                (m['codigo'],),
            ).fetchone()
            if pr_row and pr_row[0]:
                costo_total += float(pr_row[0]) * m['unidades']
        except Exception:
            pass
    return jsonify({
        'id': env.get('id'),
        'lote': env.get('lote'),
        'producto': env.get('producto'),
        'presentacion': env.get('presentacion'),
        'unidades': env.get('unidades'),
        'envase_codigo': env.get('envase_codigo'),
        'tapa_codigo': env.get('tapa_codigo'),
        'fecha': env.get('fecha'),
        'operador': env.get('operador'),
        'observaciones': env.get('observaciones', ''),
        'produccion_id': env.get('produccion_id'),
        'estado': env.get('estado', ''),
        'mee_descontados': mee_movs,
        'costo_estimado_mee_cop': round(costo_total, 2),
    })

@bp.route('/api/envasado/pendientes-acond', methods=['GET'])
def envasado_pendientes():
    """Retorna envasados Completado que no tienen acondicionamiento asociado."""
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT e.* FROM envasado e
                 LEFT JOIN acondicionamiento a ON a.envasado_id = e.id
                 WHERE a.id IS NULL
                 ORDER BY e.id DESC""")
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({'pendientes': rows})

@bp.route('/api/acondicionamiento/pendientes-lib', methods=['GET'])
def acond_pendientes_lib():
    """Retorna acondicionamientos completados sin liberacion asociada."""
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT a.id, a.lote, a.producto, a.unidades_producidas,
                        a.presentacion, a.cantidad_batch_g, a.fecha,
                        a.sku, a.precio_base
                 FROM acondicionamiento a
                 LEFT JOIN liberaciones l ON l.acondicionamiento_id = a.id
                 WHERE l.id IS NULL
                 ORDER BY a.id DESC""")
    cols = ['id','lote','producto','unidades','presentacion','batch_g','fecha','sku','precio_base']
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({'pendientes': rows})

@bp.route('/api/acondicionamiento', methods=['GET', 'POST'])
def acondicionamiento_list():
    # P0 audit 26-may-2026 · sin gate · POST creaba acondicionamiento + UPDATE MEE.
    _u_sess, _err_s, _code_s = _require_session()
    if _err_s:
        return _err_s, _code_s
    if request.method == 'POST':
        _u_w, _err_w, _code_w = _require_planta_write()
        if _err_w:
            return _err_w, _code_w
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.get_json(silent=True) or {}
        u = session.get('compras_user', '')
        mee_items = d.get('mee_consumido', [])
        envasado_id = int(d.get('envasado_id', 0) or 0)
        batch_g = float(d.get('batch_g', 0) or d.get('cantidad_batch_g', 0) or 0)
        uds = int(d.get('unidades', 0) or d.get('unidades_producidas', 0) or 0)
        c.execute("""INSERT INTO acondicionamiento
            (envasado_id, produccion_id, lote, producto, cantidad_batch_g, unidades_producidas,
             presentacion, mee_consumido, fecha, operador, observaciones, sku, precio_base)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (envasado_id, int(d.get('produccion_id') or 0), d.get('lote', ''), d.get('producto', ''),
             batch_g, uds,
             d.get('presentacion', ''), json.dumps(mee_items),
             d.get('fecha', datetime.now().strftime('%Y-%m-%d')), u,
             d.get('observaciones', ''), d.get('sku', '').strip(),
             float(d.get('precio_base', 0) or 0)))
        new_id = c.lastrowid
        # Auto-descontar MEE del maestro_mee
        lote_ref = d.get('lote', '')
        for item in mee_items:
            cod = str(item.get('codigo', item.get('codigo_mee', ''))).strip()
            cant = float(item.get('cantidad', 0) or 0)
            if not cod or cant <= 0: continue
            c.execute("SELECT stock_actual FROM maestro_mee WHERE codigo=?", (cod,))
            row = c.fetchone()
            if not row: continue
            nuevo = max(0, row[0] - cant)
            c.execute("UPDATE maestro_mee SET stock_actual=? WHERE codigo=?", (nuevo, cod))
            c.execute("""INSERT INTO movimientos_mee
                         (mee_codigo, tipo, cantidad, lote_ref, batch_ref, responsable, observaciones)
                         VALUES (?,?,?,?,?,?,?)""",
                      (cod, 'Salida', cant, lote_ref, lote_ref, u,
                       f'Consumo acondicionamiento {lote_ref}'))
        # Reemplazo MyBatch · 10-jun-2026 · LEGAJO OA AUTOMÁTICO de acondicionamiento:
        # al acondicionar nace el EBR de fase ACONDICIONAMIENTO (la "Orden de
        # Acondicionamiento") si el producto tiene MBR aprobado con pasos de OA.
        # Auto-gateado: NO_MBR_APROBADO → no-op (no bloquea · idéntico a antes para
        # productos sin MBR). Idempotente por (produccion_id, lote, fase): N
        # presentaciones del mismo lote = 1 solo legajo OA. Espeja el hook de
        # envasado (línea ~11475) y el de fabricación (línea ~2352).
        _prod_oa = (d.get('producto', '') or '').strip()
        _lote_oa = (d.get('lote', '') or '').strip()
        if _prod_oa and _lote_oa:
            try:
                from blueprints.brd import crear_ebr_desde_mbr
                _ro = crear_ebr_desde_mbr(
                    c, producto_nombre=_prod_oa, lote=_lote_oa,
                    produccion_id=(int(d.get('produccion_id') or 0) or None),
                    usuario=u, fase='acondicionamiento',
                    area_codigo=(d.get('area_codigo') or '').strip())
                if _ro.get('ok') and not _ro.get('reusado'):
                    try:
                        from database import audit_log as _alo
                        _alo(c, usuario=u or 'sistema', accion='CREAR_EBR_OA_AUTO',
                             tabla='ebr_ejecuciones', registro_id=str(_ro.get('id')),
                             despues={'producto': _prod_oa, 'lote': _lote_oa,
                                      'fase': 'acondicionamiento',
                                      'numero_op': _ro.get('numero_op')})
                    except Exception:
                        pass
            except Exception as _eo:
                __import__('logging').getLogger('inventario').warning(
                    'crear EBR OA (acondicionamiento) auto fallo (no bloquea): %s', _eo)
        conn.commit()
        return jsonify({'ok': True, 'id': new_id}), 201
    # Sprint Acondicionamiento PRO · 21-may-2026 · paginación + filtros
    try:
        limit = max(1, min(int(request.args.get('limit', 100)), 500))
    except (ValueError, TypeError):
        limit = 100
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except (ValueError, TypeError):
        offset = 0
    q = (request.args.get('q') or '').strip()
    where = ['1=1']
    params = []
    if q:
        qesc = q.replace('\\', '\\\\').replace('%', '\\%').replace('_', '\\_')
        where.append("(LOWER(COALESCE(producto,'')) LIKE LOWER(?) ESCAPE '\\' OR LOWER(COALESCE(lote,'')) LIKE LOWER(?) ESCAPE '\\' OR LOWER(COALESCE(operador,'')) LIKE LOWER(?) ESCAPE '\\')")
        params += [f'%{qesc}%', f'%{qesc}%', f'%{qesc}%']
    where_sql = ' AND '.join(where)
    try:
        total = int(c.execute(f"SELECT COUNT(*) FROM acondicionamiento WHERE {where_sql}", params).fetchone()[0] or 0)
    except Exception:
        total = 0
    c.execute(f"""SELECT id, produccion_id, lote, producto, cantidad_batch_g, unidades_producidas,
                        presentacion, fecha, operador, estado, observaciones, sku
                 FROM acondicionamiento
                 WHERE {where_sql}
                 ORDER BY creado_en DESC, id DESC
                 LIMIT ? OFFSET ?""", params + [limit, offset])
    cols = [dd[0] for dd in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify({'items': rows, 'total': total, 'limit': limit, 'offset': offset})


@bp.route('/api/acondicionamiento/<int:aid>/detalle', methods=['GET'])
def acondicionamiento_detalle(aid):
    """Sprint Acondicionamiento PRO · detalle completo del lote."""
    u, err, code = _require_session()
    if err:
        return err, code
    conn = get_db(); c = conn.cursor()
    row = c.execute("SELECT * FROM acondicionamiento WHERE id=?", (aid,)).fetchone()
    if not row:
        return jsonify({'error': 'acondicionamiento no existe'}), 404
    cols = [dd[0] for dd in c.description]
    info = dict(zip(cols, row))
    # MEE consumido (JSON)
    try:
        import json as _j
        info['mee_consumido_parsed'] = _j.loads(info.get('mee_consumido') or '[]')
    except Exception:
        info['mee_consumido_parsed'] = []
    return jsonify(info)

@bp.route('/api/acondicionamiento/<int:aid>', methods=['PATCH'])
def acondicionamiento_update(aid):
    u, err, code = _require_planta_write()
    if err:
        return err, code
    d = request.get_json(silent=True) or {}
    conn = get_db(); c = conn.cursor()
    antes_row = c.execute(
        "SELECT estado, unidades_producidas, observaciones FROM acondicionamiento WHERE id=?",
        (aid,)).fetchone()
    if not antes_row:
        return jsonify({'error': 'Acondicionamiento no encontrado'}), 404
    antes = dict(antes_row)
    if 'estado' in d: c.execute("UPDATE acondicionamiento SET estado=? WHERE id=?", (d['estado'], aid))
    if 'unidades_producidas' in d: c.execute("UPDATE acondicionamiento SET unidades_producidas=? WHERE id=?", (int(d['unidades_producidas']), aid))
    if 'mee_consumido' in d: c.execute("UPDATE acondicionamiento SET mee_consumido=? WHERE id=?", (json.dumps(d['mee_consumido']), aid))
    if 'observaciones' in d: c.execute("UPDATE acondicionamiento SET observaciones=? WHERE id=?", (d['observaciones'], aid))
    audit_log(c, usuario=u, accion='ACTUALIZAR_ACONDICIONAMIENTO',
              tabla='acondicionamiento', registro_id=aid, antes=antes,
              despues={k: d.get(k) for k in ('estado','unidades_producidas','observaciones') if k in d},
              detalle=f"Actualizó acondicionamiento id={aid}"
                      + (f" · estado→{d['estado']}" if 'estado' in d else ""))
    conn.commit()
    return jsonify({'ok': True})

@bp.route('/api/liberacion', methods=['GET', 'POST'])
def liberacion_list():
    # P0 audit 26-may-2026 · liberación PT es decisión INVIMA art. 10 ·
    # historicamente este endpoint NO tenía gate · cualquier anónimo
    # podía INSERT en `liberaciones` (registro regulatorio).
    if request.method == 'POST':
        u_qc, err_qc, code_qc = _require_qc()
        if err_qc:
            return err_qc, code_qc
        u = u_qc
    else:
        u_sess, err_s, code_s = _require_session()
        if err_s:
            return err_s, code_s
    conn = get_db(); c = conn.cursor()
    if request.method == 'POST':
        d = request.get_json(silent=True) or {}
        # Validate_money en precio_base · NaN/Inf permitiría contaminar histórico
        try:
            _pb_raw = d.get('precio_base', 0) or 0
            precio_base = float(_pb_raw)
            import math as _m
            if not _m.isfinite(precio_base) or precio_base < 0 or precio_base > 1e10:
                return jsonify({'error': 'precio_base inválido'}), 400
        except (TypeError, ValueError):
            return jsonify({'error': 'precio_base no numérico'}), 400
        try:
            unidades = int(d.get('unidades', 0))
        except (TypeError, ValueError):
            return jsonify({'error': 'unidades inválido'}), 400
        if unidades <= 0:
            return jsonify({'error': 'unidades debe ser > 0'}), 400
        c.execute("""INSERT INTO liberaciones
            (acondicionamiento_id, lote, producto, unidades, presentacion,
             fecha_produccion, cliente, destino, observaciones, sku, precio_base)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (int(d.get('acondicionamiento_id', 0)), d.get('lote', ''), d.get('producto', ''),
             unidades, d.get('presentacion', ''), d.get('fecha_produccion', ''),
             d.get('cliente', ''), d.get('destino', 'ANIMUS'), d.get('observaciones', ''),
             (d.get('sku', '') or '').strip(), precio_base))
        new_id = c.lastrowid
        # Audit log regulatorio INVIMA art. 10
        try:
            from audit_helpers import audit_log as _al
            _al(c, usuario=u, accion='CREAR_LIBERACION', tabla='liberaciones',
                registro_id=new_id,
                despues={'lote': d.get('lote',''), 'sku': (d.get('sku','') or '').strip(),
                         'unidades': unidades, 'cliente': d.get('cliente',''),
                         'destino': d.get('destino','ANIMUS')},
                detalle=f"Liberación lote={d.get('lote','')} sku={(d.get('sku','') or '').strip()} u={unidades}")
        except Exception as _ae:
            import logging as _lg
            _lg.getLogger('inventario').warning('audit liberacion fallo: %s', _ae)
        conn.commit()
        return jsonify({'ok': True, 'id': new_id}), 201
    estado = request.args.get('estado', '')
    if estado: c.execute("SELECT * FROM liberaciones WHERE estado=? ORDER BY creado_en DESC LIMIT 100", (estado,))
    else: c.execute("SELECT * FROM liberaciones ORDER BY creado_en DESC LIMIT 100")
    cols = [d[0] for d in c.description]
    rows = [dict(zip(cols, r)) for r in c.fetchall()]
    return jsonify(rows)

@bp.route('/api/liberacion/<int:lid>', methods=['PATCH'])
def liberacion_update(lid):
    """Aprueba/rechaza liberación de PT · DECISIÓN INVIMA art. 10.

    Solo Calidad/Admin pueden Liberar/Rechazar. Rechazo requiere observaciones
    (≥10 chars) como motivo regulatorio.
    """
    # RBAC: solo Calidad/Admin para decisiones Liberado/Rechazado
    u_qc, err_qc, code_qc = _require_qc()
    if err_qc:
        return err_qc, code_qc
    u = u_qc
    d = request.get_json(silent=True) or {}
    estado = d.get('estado', '')
    obs = (d.get('observaciones') or '').strip()
    if estado == 'Rechazado' and len(obs) < 10:
        return jsonify({'error': 'observaciones (≥10 chars) requeridas para rechazar liberación'}), 400
    conn = get_db(); c = conn.cursor()
    antes_row = c.execute(
        "SELECT lote, producto, unidades, sku, estado FROM liberaciones WHERE id=?",
        (lid,)).fetchone()
    if not antes_row:
        return jsonify({'error': 'Liberación no encontrada'}), 404
    antes = dict(antes_row)
    if estado == 'Liberado':
        c.execute("UPDATE liberaciones SET estado='Liberado', fecha_liberacion=?, aprobado_por=?, cliente=? WHERE id=?",
                  (datetime.now().strftime('%Y-%m-%d'), u, d.get('cliente', ''), lid))
        # Auto-crear entrada en stock_pt al liberar
        c.execute("SELECT lote, producto, unidades, sku, precio_base, presentacion FROM liberaciones WHERE id=?", (lid,))
        lib = c.fetchone()
        if lib and lib[3]:  # sku presente
            lote_lib, prod_lib, uds_lib, sku_lib, precio_lib, pres_lib = lib
            fecha_lib = datetime.now().strftime('%Y-%m-%d')
            c.execute("""INSERT INTO stock_pt
                         (sku, descripcion, lote_produccion, fecha_produccion,
                          unidades_inicial, unidades_disponible, precio_base,
                          empresa, estado, observaciones)
                         VALUES (?,?,?,?,?,?,?,?,?,?)""",
                      (sku_lib, prod_lib, lote_lib, fecha_lib,
                       uds_lib, uds_lib, precio_lib,
                       'ANIMUS', 'Disponible',
                       f'Liberacion aprobada por {u} — {pres_lib}'))
        # Registrar en calidad como BPM completado
        try:
            _lib_val = lib[1] if lib else 'PT'
            _lib_lote = lib[0] if lib else ''
            _cli_dest = d.get('cliente','') or 'sin cliente'
            c.execute("""INSERT INTO calidad_registros
                         (fecha, tarea_id, usuario, estado, valor_registrado, observaciones)
                         VALUES (date('now', '-5 hours'), NULL, ?, 'Completado', ?, ?)""",
                     (u,
                      f"{_lib_lote} | {str(_lib_val)[:40]}",
                      f"BPM Liberacion PT -> {_cli_dest}"))
        except Exception:
            pass
    elif estado == 'Rechazado':
        c.execute("UPDATE liberaciones SET estado='Rechazado', aprobado_por=?, observaciones=? WHERE id=?",
                  (u, obs, lid))
    else:
        c.execute("UPDATE liberaciones SET observaciones=? WHERE id=?", (obs, lid))
    accion = 'LIBERAR_PT' if estado == 'Liberado' else 'RECHAZAR_PT' if estado == 'Rechazado' else 'ACTUALIZAR_LIBERACION'
    audit_log(c, usuario=u, accion=accion, tabla='liberaciones',
              registro_id=lid, antes=antes,
              despues={'estado': estado, 'aprobado_por': u, 'observaciones': obs,
                       'cliente': d.get('cliente', '')},
              detalle=f"{accion} lote {antes.get('lote','—')} ({antes.get('producto','')})"
                      + (f" · {obs}" if obs else ""))
    conn.commit()
    return jsonify({'ok': True})


# ─── ALERTAS VIVAS DE PLANTA ─────────────────────────────────────────────────
# Endpoint unificado que el panel admin / centro de mando consume para mostrar
# todo lo que requiere atención HOY. Cubre items #3, #4, #6 del audit:
#   - Vencimientos próximos (<30 días) y vencidos
#   - Stock por debajo de mínimo
#   - Conteos cíclicos cerrados con discrepancias > tolerancia, sin ajuste
#   - Lotes en cuarentena que llevan > 5 días esperando QC

@bp.route('/api/planta/alertas-vivas', methods=['GET'])
def alertas_vivas_planta():
    """Consolida todas las alertas operacionales de Planta en un solo endpoint.

    Retorna { vencimientos, stock_bajo, discrepancias, cuarentena_extendida,
              total, severidad_max }
    severidad_max: 'critico' | 'alto' | 'medio' | 'ok'
    """
    u, err, code = _require_session()
    if err:
        return err, code

    conn = get_db(); c = conn.cursor()
    from datetime import date, timedelta
    hoy = date.today()
    in_30 = (hoy + timedelta(days=30)).isoformat()
    hace_5 = (hoy - timedelta(days=5)).isoformat()

    # ── 1. Vencimientos ───────────────────────────────────────────────────────
    # MPs con stock vivo cuyo lote vence en <30 días (o ya vencido)
    c.execute("""
        SELECT m.material_id, m.material_nombre, m.lote, m.fecha_vencimiento,
               COALESCE(mp.tipo_material,'MP') as tipo,
               SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) as stock
        FROM movimientos m
        LEFT JOIN maestro_mps mp ON m.material_id = mp.codigo_mp
        WHERE m.fecha_vencimiento != ''
          AND m.fecha_vencimiento <= ?
          AND m.estado_lote IN ('VIGENTE','CUARENTENA')
        -- material_nombre/fecha_vencimiento/tipo son funcionalmente dependientes
        -- del lote, pero PG exige toda columna no-agregada en el GROUP BY (SQLite
        -- no) → si no, 500 'must appear in the GROUP BY clause'. Y PG tampoco
        -- acepta el alias 'stock' en HAVING → repetir la expresión. Cazado por suite PG.
        GROUP BY m.material_id, m.lote, m.material_nombre, m.fecha_vencimiento, mp.tipo_material
        HAVING SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END) > 0
        ORDER BY m.fecha_vencimiento ASC
        LIMIT 100
    """, (in_30,))
    venc_rows = c.fetchall()
    vencimientos = []
    for r in venc_rows:
        try:
            fv = date.fromisoformat(r[3][:10])
            dias = (fv - hoy).days
        except (ValueError, TypeError, IndexError):
            dias = None
        vencimientos.append({
            'material_id': r[0], 'material_nombre': r[1], 'lote': r[2],
            'fecha_vencimiento': r[3], 'dias_restantes': dias,
            'tipo_material': r[4], 'stock_g': round(r[5] or 0, 1),
            'severidad': 'critico' if dias is not None and dias < 0 else
                         ('alto' if dias is not None and dias <= 7 else 'medio')
        })

    # ── 2. Stock bajo mínimo ──────────────────────────────────────────────────
    c.execute("""
        SELECT mp.codigo_mp, mp.nombre_comercial, mp.stock_minimo,
               COALESCE(mp.tipo_material,'MP') as tipo,
               COALESCE((
                   SELECT SUM(CASE WHEN m.tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN m.cantidad WHEN m.tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -m.cantidad ELSE 0 END)
                   FROM movimientos m
                   WHERE m.material_id = mp.codigo_mp AND m.estado_lote='VIGENTE'
               ), 0) as stock_actual
        FROM maestro_mps mp
        WHERE mp.activo = 1 AND COALESCE(mp.stock_minimo, 0) > 0
    """)
    stock_bajo = []
    for r in c.fetchall():
        codigo, nombre, st_min, tipo, st_act = r
        if st_act < st_min:
            ratio = st_act / st_min if st_min else 0
            stock_bajo.append({
                'codigo_mp': codigo, 'nombre': nombre,
                'stock_minimo': st_min, 'stock_actual': round(st_act, 1),
                'tipo_material': tipo, 'cobertura_pct': round(ratio * 100, 1),
                'severidad': 'critico' if ratio < 0.25 else ('alto' if ratio < 0.5 else 'medio')
            })
    stock_bajo.sort(key=lambda x: x['cobertura_pct'])

    # ── 3. Discrepancias de conteo cíclico no resueltas ───────────────────────
    discrepancias = []
    try:
        # ajuste_aplicado vive en conteo_items, NO en conteos_fisicos.
        # La fecha de cierre es fecha_cierre. Un conteo está "no resuelto"
        # si tiene items con diferencia != 0 y ajuste_aplicado = 0.
        c.execute("""
            SELECT cf.id, cf.numero, cf.fecha_cierre,
                   COUNT(ci.id) as n_items,
                   SUM(CASE WHEN ABS(COALESCE(ci.diferencia,0)) > 0
                             AND COALESCE(ci.ajuste_aplicado,0) = 0
                            THEN 1 ELSE 0 END) as n_dif
            FROM conteos_fisicos cf
            LEFT JOIN conteo_items ci ON cf.id = ci.conteo_id
            WHERE cf.estado = 'Cerrado'
            -- PG: numero/fecha_cierre deben ir en GROUP BY, y HAVING no acepta
            -- el alias n_dif → repetir la expresión. Si esta query falla en PG la
            -- transacción queda abortada y revienta TODO el endpoint (el except no
            -- la recupera). Cazado por suite PG.
            GROUP BY cf.id, cf.numero, cf.fecha_cierre
            HAVING SUM(CASE WHEN ABS(COALESCE(ci.diferencia,0)) > 0
                             AND COALESCE(ci.ajuste_aplicado,0) = 0
                            THEN 1 ELSE 0 END) > 0
            ORDER BY cf.fecha_cierre DESC
            LIMIT 20
        """)
        for r in c.fetchall():
            discrepancias.append({
                'conteo_id': r[0], 'numero': r[1], 'estanteria': '',
                'cerrado_en': r[2], 'items_con_diferencia': r[4],
                'severidad': 'alto'
            })
    except sqlite3.OperationalError:
        # Red de seguridad ante esquema legacy · no debería dispararse
        pass

    # ── 4. Cuarentenas extendidas (>5 días esperando QC) ──────────────────────
    c.execute("""
        SELECT m.id, m.material_id, m.material_nombre, m.lote, m.fecha,
               m.cantidad, m.proveedor
        FROM movimientos m
        WHERE m.estado_lote IN ('CUARENTENA', 'CUARENTENA_EXTENDIDA')
          AND m.tipo = 'Entrada'
          AND m.fecha < ?
        ORDER BY m.fecha ASC
        LIMIT 50
    """, (hace_5,))
    cuarentena_extendida = [
        {'mov_id': r[0], 'material_id': r[1], 'material_nombre': r[2],
         'lote': r[3], 'fecha_ingreso': r[4], 'cantidad_g': r[5],
         'proveedor': r[6], 'severidad': 'alto'}
        for r in c.fetchall()
    ]

    total = (len(vencimientos) + len(stock_bajo) +
             len(discrepancias) + len(cuarentena_extendida))
    sevs = (
        [v['severidad'] for v in vencimientos] +
        [v['severidad'] for v in stock_bajo] +
        [v['severidad'] for v in discrepancias] +
        [v['severidad'] for v in cuarentena_extendida]
    )
    sev_orden = {'critico': 3, 'alto': 2, 'medio': 1, 'ok': 0}
    severidad_max = max(sevs, key=lambda s: sev_orden.get(s, 0)) if sevs else 'ok'

    return jsonify({
        'vencimientos':           vencimientos,
        'stock_bajo':             stock_bajo,
        'discrepancias':          discrepancias,
        'cuarentena_extendida':   cuarentena_extendida,
        'total':                  total,
        'severidad_max':          severidad_max,
        'evaluado_en':            hoy.isoformat(),
    })


# ─── KARDEX + VALORACIÓN FIFO ────────────────────────────────────────────────
# Reporte estándar contable para auditoría y costeo.

@bp.route('/api/planta/kardex/<codigo_mp>', methods=['GET'])
def planta_kardex(codigo_mp):
    """Kardex de un MP específico: entradas, salidas, saldo running, valor FIFO.

    Query params:
      - desde: YYYY-MM-DD (opcional, default hace 12 meses)
      - hasta: YYYY-MM-DD (opcional, default hoy)
    """
    u, err, code = _require_session()
    if err:
        return err, code

    from datetime import date, timedelta
    desde = (request.args.get('desde') or
             (date.today() - timedelta(days=365)).isoformat())
    hasta = request.args.get('hasta') or date.today().isoformat()

    conn = get_db(); c = conn.cursor()

    # Datos maestros
    c.execute("""SELECT codigo_mp, nombre_comercial, nombre_inci,
                        COALESCE(precio_referencia,0), COALESCE(stock_minimo,0),
                        COALESCE(tipo_material,'MP')
                 FROM maestro_mps WHERE codigo_mp=?""", (codigo_mp,))
    mp = c.fetchone()
    if not mp:
        return jsonify({'error': f'MP {codigo_mp} no existe'}), 404

    # Movimientos del rango
    c.execute("""SELECT id, fecha, tipo, cantidad, lote, observaciones,
                        proveedor, COALESCE(precio_kg,0), numero_oc, numero_factura,
                        estado_lote, fecha_vencimiento
                 FROM movimientos
                 WHERE material_id=? AND fecha >= ? AND fecha <= ?
                 ORDER BY fecha ASC, id ASC""",
              (codigo_mp, desde, hasta + 'T23:59:59'))
    rows = c.fetchall()

    # Algoritmo FIFO: cada entrada va a una "capa" (queue) con su precio.
    # Cada salida consume capas FIFO (más antigua primero) y registra el
    # costo unitario ponderado de esa salida.
    capas = []  # [(cantidad_kg_restante, precio_kg, lote)]
    movimientos = []
    saldo_g = 0.0
    valor_acum = 0.0

    for r in rows:
        mid, fecha, tipo, cant, lote, obs, prov, pkg, oc, fac, est_lote, fvenc = r
        cant_kg = (cant or 0) / 1000.0

        if tipo == 'Entrada':
            saldo_g += (cant or 0)
            costo_entrada = cant_kg * (pkg or 0)
            valor_acum += costo_entrada
            capas.append({
                'cantidad_kg_restante': cant_kg,
                'precio_kg': pkg or 0,
                'lote': lote, 'fecha': fecha,
            })
            mov = {
                'id': mid, 'fecha': fecha, 'tipo': 'Entrada',
                'cantidad_kg': round(cant_kg, 3), 'precio_kg': round(pkg or 0, 2),
                'costo': round(costo_entrada, 2), 'lote': lote, 'proveedor': prov,
                'oc': oc, 'factura': fac, 'estado_lote': est_lote,
                'fecha_vencimiento': fvenc,
                'saldo_g_running': round(saldo_g, 1),
                'valor_running': round(valor_acum, 2),
            }
        else:  # Salida
            saldo_g -= (cant or 0)
            # Consumir capas FIFO
            falta_kg = cant_kg
            costo_salida = 0.0
            consumos = []
            i = 0
            while falta_kg > 1e-9 and i < len(capas):
                if capas[i]['cantidad_kg_restante'] <= 1e-9:
                    i += 1
                    continue
                tomar = min(capas[i]['cantidad_kg_restante'], falta_kg)
                costo_salida += tomar * capas[i]['precio_kg']
                consumos.append({
                    'lote': capas[i]['lote'],
                    'kg': round(tomar, 3),
                    'precio_kg': capas[i]['precio_kg']
                })
                capas[i]['cantidad_kg_restante'] -= tomar
                falta_kg -= tomar
                if capas[i]['cantidad_kg_restante'] <= 1e-9:
                    i += 1
            valor_acum -= costo_salida
            costo_unit = (costo_salida / cant_kg) if cant_kg > 0 else 0
            mov = {
                'id': mid, 'fecha': fecha, 'tipo': 'Salida',
                'cantidad_kg': round(cant_kg, 3),
                'costo_unit_kg': round(costo_unit, 2),
                'costo_total': round(costo_salida, 2),
                'lote': lote, 'observaciones': obs,
                'consumos_fifo': consumos,
                'saldo_g_running': round(saldo_g, 1),
                'valor_running': round(valor_acum, 2),
            }
        movimientos.append(mov)

    # Capas remanentes = stock actual valorado FIFO
    stock_actual = [
        {
            'lote': cp['lote'],
            'cantidad_kg': round(cp['cantidad_kg_restante'], 3),
            'precio_kg': cp['precio_kg'],
            'valor': round(cp['cantidad_kg_restante'] * cp['precio_kg'], 2),
            'fecha_entrada': cp['fecha'],
        }
        for cp in capas if cp['cantidad_kg_restante'] > 1e-6
    ]

    # Totales
    total_entradas = sum(m['cantidad_kg'] for m in movimientos if m['tipo'] == 'Entrada')
    total_salidas = sum(m['cantidad_kg'] for m in movimientos if m['tipo'] == 'Salida')
    valor_actual = round(sum(c['valor'] for c in stock_actual), 2)

    return jsonify({
        'mp': {
            'codigo_mp': mp[0], 'nombre_comercial': mp[1],
            'nombre_inci': mp[2], 'precio_referencia': mp[3],
            'stock_minimo': mp[4], 'tipo_material': mp[5],
        },
        'rango': {'desde': desde, 'hasta': hasta},
        'totales': {
            'entradas_kg': round(total_entradas, 3),
            'salidas_kg':  round(total_salidas, 3),
            'saldo_actual_g': round(saldo_g, 1),
            'valor_actual_fifo': valor_actual,
            'movimientos_count': len(movimientos),
        },
        'movimientos': movimientos,
        'stock_actual_capas': stock_actual,
    })


@bp.route('/api/planta/stock-por-lote/<codigo_mp>', methods=['GET'])
def planta_stock_por_lote(codigo_mp):
    """Stock detallado por lote de un MP (FEFO view).

    Sebastian (29-abr-2026): "FEFO perfecto". Endpoint que muestra el
    stock vivo por lote — cuántos gramos quedan de cada lote, ordenado
    por fecha de vencimiento (más cercano primero). Útil para que
    operarios sepan exactamente de qué lote sacar al consumir.

    Returns: {
        codigo_mp, nombre,
        lotes: [{lote, fecha_vencimiento, estado, stock_g, dias_a_vencer}, ...],
        sin_lote_g: <stock que entró sin lote>,
        total_g: <stock total>
    }
    """
    if 'compras_user' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    conn = get_db(); c = conn.cursor()

    mp = c.execute(
        "SELECT codigo_mp, COALESCE(nombre_comercial, nombre_inci, codigo_mp) "
        "FROM maestro_mps WHERE codigo_mp=?", (codigo_mp,)
    ).fetchone()
    if not mp:
        return jsonify({'error': f'MP {codigo_mp} no existe'}), 404

    # Stock por lote (Entradas - Salidas con lote)
    rows = c.execute("""
        -- FIX-B1 12-may-2026: GROUP BY solo `lote`. Antes agrupaba por
        -- (lote, fecha_vencimiento, estado_lote) lo que separaba las
        -- Entradas (que llenan fv/estado) de las Salidas FEFO de
        -- _distribuir_fefo (que no las llenan, quedan NULL). Resultado:
        -- el grupo Salida quedaba con stock_g negativo y HAVING > 0 lo
        -- filtraba, dejando el grupo Entrada con stock viejo. UI Bodega
        -- mostraba 1000g cuando lo real eran 500g.
        SELECT lote,
               MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END) as fecha_vencimiento,
               MAX(CASE WHEN tipo='Entrada' THEN estado_lote END) as estado_lote,
               SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END) as stock_g,
               MIN(fecha) as primera_entrada
        FROM movimientos
        WHERE material_id = ?
          AND COALESCE(lote,'') != ''
        GROUP BY lote
        HAVING stock_g > 0
        ORDER BY COALESCE(MAX(CASE WHEN tipo='Entrada' THEN fecha_vencimiento END), '9999-12-31') ASC, lote ASC
    """, (codigo_mp,)).fetchall()

    from datetime import date
    hoy = date.today()
    lotes = []
    total_con_lote = 0.0
    for lote, fv, estado, stock_g, primera in rows:
        dias_a_vencer = None
        if fv:
            try:
                fv_date = date.fromisoformat(fv[:10])
                dias_a_vencer = (fv_date - hoy).days
            except Exception:
                pass
        lotes.append({
            'lote': lote,
            'fecha_vencimiento': fv,
            'estado_lote': estado or 'OK',
            'stock_g': round(float(stock_g), 2),
            'dias_a_vencer': dias_a_vencer,
            'primera_entrada': primera,
        })
        total_con_lote += float(stock_g)

    # Stock sin lote (entradas legacy / sin trazabilidad)
    sin_lote_row = c.execute("""
        SELECT COALESCE(SUM(CASE WHEN tipo IN ('Entrada','entrada','ENTRADA','Ajuste +','Ajuste') THEN cantidad WHEN tipo IN ('Salida','salida','SALIDA','Ajuste -') THEN -cantidad ELSE 0 END), 0)
        FROM movimientos
        WHERE material_id = ?
          AND COALESCE(lote,'') = ''
    """, (codigo_mp,)).fetchone()
    sin_lote_g = round(float(sin_lote_row[0] or 0), 2)

    total_g = round(total_con_lote + sin_lote_g, 2)
    return jsonify({
        'codigo_mp': mp[0],
        'nombre': mp[1],
        'lotes': lotes,
        'sin_lote_g': sin_lote_g,
        'total_g': total_g,
        'total_lotes_activos': len(lotes),
    })


@bp.route('/api/planta/valoracion-inventario', methods=['GET'])
def planta_valoracion_inventario():
    """Valoración total del inventario FIFO de todas las MPs activas.

    Calcula el valor de cada MP con su stock actual al precio FIFO
    (capas en orden de ingreso). Útil para reporte contable y cierre.
    Query: ?tipo_material=MP|Envase Primario|Envase Secundario|Empaque
    """
    u, err, code = _require_session()
    if err:
        return err, code

    tipo_filter = (request.args.get('tipo_material') or '').strip()
    # FIX · 21-may-2026 · whitelist estricta (antes valores fuera lista colapsaban a 'TODOS')
    TIPOS_VALIDOS = ('MP', 'Envase Primario', 'Envase Secundario', 'Empaque')
    if tipo_filter and tipo_filter not in TIPOS_VALIDOS:
        return jsonify({
            'error': f'tipo_material inválido · usar {TIPOS_VALIDOS} o vacío',
            'codigo': 'TIPO_MATERIAL_INVALIDO',
        }), 400
    conn = get_db(); c = conn.cursor()

    # MPs candidatas
    if tipo_filter in TIPOS_VALIDOS:
        c.execute("""SELECT codigo_mp, nombre_comercial, COALESCE(tipo_material,'MP')
                     FROM maestro_mps WHERE activo=1 AND tipo_material=?""",
                  (tipo_filter,))
    else:
        c.execute("""SELECT codigo_mp, nombre_comercial, COALESCE(tipo_material,'MP')
                     FROM maestro_mps WHERE activo=1""")
    mps = c.fetchall()

    out = []
    valor_total = 0.0
    for codigo, nombre, tipo in mps:
        c.execute("""SELECT fecha, tipo, cantidad, COALESCE(precio_kg,0), lote
                     FROM movimientos WHERE material_id=?
                     ORDER BY fecha ASC, id ASC""", (codigo,))
        movs = c.fetchall()
        capas = []
        stock_g = 0
        for mv in movs:
            fecha, tp, cant, pkg, lote = mv
            cant_kg = (cant or 0) / 1000.0
            if tp == 'Entrada':
                stock_g += (cant or 0)
                capas.append([cant_kg, pkg or 0, lote])
            else:
                stock_g -= (cant or 0)
                falta = cant_kg
                for capa in capas:
                    if falta <= 1e-9:
                        break
                    if capa[0] <= 1e-9:
                        continue
                    tomar = min(capa[0], falta)
                    capa[0] -= tomar
                    falta -= tomar
        valor = sum(cap[0] * cap[1] for cap in capas if cap[0] > 1e-6)
        if stock_g > 0 or valor > 0:
            out.append({
                'codigo_mp': codigo, 'nombre': nombre, 'tipo_material': tipo,
                'stock_g': round(stock_g, 1),
                'stock_kg': round(stock_g / 1000, 3),
                'valor_fifo': round(valor, 2),
            })
            valor_total += valor

    out.sort(key=lambda x: x['valor_fifo'], reverse=True)

    # Totales por tipo de material
    por_tipo = {}
    for item in out:
        t = item['tipo_material']
        por_tipo.setdefault(t, {'count': 0, 'valor': 0, 'stock_kg': 0})
        por_tipo[t]['count'] += 1
        por_tipo[t]['valor'] += item['valor_fifo']
        por_tipo[t]['stock_kg'] += item['stock_kg']
    for t in por_tipo:
        por_tipo[t]['valor'] = round(por_tipo[t]['valor'], 2)
        por_tipo[t]['stock_kg'] = round(por_tipo[t]['stock_kg'], 3)

    return jsonify({
        'items': out,
        'valor_total_fifo': round(valor_total, 2),
        'count': len(out),
        'por_tipo_material': por_tipo,
        'filtro_tipo': tipo_filter or 'todos',
    })


# ═══════════════════════════════════════════════
#  MAQUILA 360 — API
# ═══════════════════════════════════════════════
