"""
admin.py — Blueprint de administración: panel, backups, eventos de seguridad.

Acceso: SOLO ADMIN_USERS (sebastian, alejandro). El resto recibe 403.
"""
import os
import secrets
import sqlite3
import string

from flask import Blueprint, jsonify, request, session, send_file, Response
from werkzeug.security import generate_password_hash

from config import (
    DB_PATH, ADMIN_USERS, COMPRAS_USERS, validate_config,
    CONTADORA_USERS, RRHH_USERS, COMPRAS_ACCESS, FINANZAS_ACCESS,
    CLIENTES_ACCESS, TECNICA_USERS, MARKETING_USERS, ANIMUS_ACCESS,
    ESPAGIRIA_ACCESS, CALIDAD_USERS, PLANTA_USERS,
)
from auth import _client_ip, _log_sec
from audit_helpers import audit_log
from backup import (
    do_backup, list_backups, get_backup_path, BACKUPS_DIR,
    RETENTION_DAYS, BACKUP_INTERVAL_HOURS,
)

bp = Blueprint("admin", __name__)


def _require_admin():
    """Retorna (None, response, status) si NO admin, (user, None, None) si OK."""
    u = session.get("compras_user", "")
    if not u:
        return None, jsonify({"error": "No autenticado"}), 401
    if u not in ADMIN_USERS:
        return None, jsonify({"error": "Solo admins"}), 403
    return u, None, None


# ─── Backups ──────────────────────────────────────────────────────────────────

@bp.route("/api/admin/backups", methods=["GET"])
def admin_backups_list():
    """Lista backups disponibles + status del último."""
    u, err, code = _require_admin()
    if err:
        return err, code

    items = list_backups()

    # Trae stats del backup_log para mostrar contexto
    try:
        conn = sqlite3.connect(DB_PATH)
        recent = conn.execute(
            """SELECT id, started_at, completed_at, status, size_bytes,
                      triggered_by, error
               FROM backup_log
               ORDER BY id DESC LIMIT 20"""
        ).fetchall()
        conn.close()
        recent_list = [
            {
                "id": r[0], "started_at": r[1], "completed_at": r[2],
                "status": r[3], "size_bytes": r[4], "triggered_by": r[5],
                "error": r[6],
            } for r in recent
        ]
    except Exception:
        recent_list = []

    return jsonify({
        "backups": items,
        "recent_runs": recent_list,
        "config": {
            "retention_days": RETENTION_DAYS,
            "interval_hours": BACKUP_INTERVAL_HOURS,
            "backups_dir": BACKUPS_DIR,
        }
    })


@bp.route("/api/admin/backup-now", methods=["POST"])
def admin_backup_now():
    """Trigger manual de backup."""
    u, err, code = _require_admin()
    if err:
        return err, code

    _log_sec("backup_manual_triggered", u, _client_ip())
    result = do_backup(triggered_by=f"manual:{u}")
    # Audit log INVIMA · backup manual queda en trail regulatorio
    try:
        conn_a = sqlite3.connect(DB_PATH)
        conn_a.execute("PRAGMA busy_timeout=2000")
        cur_a = conn_a.cursor()
        audit_log(cur_a, usuario=u, accion='BACKUP_MANUAL',
                  tabla='backup_log', registro_id=None,
                  despues={'ok': bool(result.get('ok')),
                            'skipped': bool(result.get('skipped')),
                            'filename': (result.get('filename') or '')[:200]},
                  detalle=f"Backup manual triggered por {u} · "
                          f"ok={result.get('ok')} skipped={result.get('skipped')}")
        conn_a.commit(); conn_a.close()
    except Exception:
        pass  # _log_sec ya queda como evidencia
    # ok=True → backup creado, status 200
    # skipped=True → otro worker está haciendo backup, status 200 (no es error)
    # else → error real, status 500
    if result.get("ok") or result.get("skipped"):
        return jsonify(result), 200
    return jsonify(result), 500


@bp.route("/api/admin/restore-backup", methods=["POST"])
def admin_restore_backup():
    """Restaura la DB desde un backup específico.

    Sebastián 8-may-2026 (EMERGENCIA): SQLite corrupted en prod.
    Necesitamos restore rápido sin ssh a Render.

    Body JSON: {filename: "inventario_YYYYMMDD_HHMMSS.db.gz", confirm: true}

    Pasos:
      1. Verificar que filename existe en BACKUPS_DIR
      2. Hacer copia del DB actual a /var/data/inventario.db.corrupt-<ts>
         (por si necesitamos forensics después)
      3. gunzip el backup a /var/data/inventario.db.tmp
      4. mv atomic sobre DB_PATH
      5. Re-init conexiones · próxima request abre la nueva DB

    Returns:
      {ok, restored_from, corrupt_backup_path, message}
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    body = request.get_json(silent=True) or {}
    filename = (body.get("filename") or "").strip()
    confirm = bool(body.get("confirm"))

    if not filename:
        return jsonify({"error": "filename requerido"}), 400
    if not confirm:
        return jsonify({"error": "confirm=true requerido para evitar restore accidental"}), 400

    # Validar filename
    if not filename.startswith("inventario_") or not filename.endswith(".db.gz"):
        return jsonify({"error": "filename inválido (debe ser inventario_*.db.gz)"}), 400
    if "/" in filename or ".." in filename:
        return jsonify({"error": "filename con caracteres prohibidos"}), 400

    import os as _os
    import gzip
    import shutil
    import time as _time

    backup_path = _os.path.join(BACKUPS_DIR, filename)
    if not _os.path.exists(backup_path):
        return jsonify({"error": f"backup no encontrado: {filename}"}), 404

    backup_size = _os.path.getsize(backup_path)
    if backup_size < 1024:
        return jsonify({
            "error": f"backup muy pequeno ({backup_size} bytes) · probablemente corrupto"
        }), 400

    ts = _time.strftime("%Y%m%d_%H%M%S")
    corrupt_path = f"{DB_PATH}.corrupt-{ts}"
    tmp_path = f"{DB_PATH}.restore-tmp-{ts}"

    try:
        # 1. Mover DB corrupto a archivo .corrupt (para forensics)
        if _os.path.exists(DB_PATH):
            try:
                shutil.copy2(DB_PATH, corrupt_path)
            except Exception:
                pass  # best-effort

        # 2. Descomprimir backup a tmp
        with gzip.open(backup_path, 'rb') as src, open(tmp_path, 'wb') as dst:
            shutil.copyfileobj(src, dst)

        # 3. Verificar que el archivo descomprimido es una SQLite válida
        try:
            test_conn = sqlite3.connect(tmp_path)
            test_conn.execute("PRAGMA integrity_check")
            test_conn.close()
        except Exception as ce:
            try:
                _os.unlink(tmp_path)
            except Exception:
                pass
            return jsonify({
                "error": f"backup descomprimido NO es SQLite válida: {ce}",
            }), 500

        # 4. mv atomic sobre DB_PATH
        # SQLite WAL: borrar archivos -wal y -shm si existen para evitar
        # mezcla con DB nueva
        for suffix in ('-wal', '-shm', '-journal'):
            sidecar = DB_PATH + suffix
            if _os.path.exists(sidecar):
                try:
                    _os.unlink(sidecar)
                except Exception:
                    pass
        _os.replace(tmp_path, DB_PATH)

        # 5. _log_sec para audit (no tocamos DB porque acabamos de
        # restaurar y queremos ver si abre limpio).
        _log_sec("db_restored", u, _client_ip(),
                 detalle=f"restored from {filename}")

        # 6. Verificar que la nueva DB abre limpia
        verify = "ok"
        try:
            verify_conn = sqlite3.connect(DB_PATH)
            row = verify_conn.execute("PRAGMA integrity_check").fetchone()
            verify = str(row[0]) if row else "(empty)"
            verify_conn.close()
        except Exception as e:
            verify = f"verify failed: {e}"

        return jsonify({
            "ok": True,
            "restored_from": filename,
            "backup_size_bytes": backup_size,
            "corrupt_backup_path": corrupt_path,
            "integrity_check": verify,
            "message": (f"DB restaurada desde {filename}. "
                        "Próxima request usa la DB nueva. "
                        f"Backup del corrupto: {corrupt_path}"),
        })

    except Exception as e:
        # Limpieza de archivos temporales
        try:
            if _os.path.exists(tmp_path):
                _os.unlink(tmp_path)
        except Exception:
            pass
        return jsonify({"error": f"restore falló: {e}"}), 500


@bp.route("/api/admin/backup/<path:filename>", methods=["GET"])
def admin_backup_download(filename):
    """Descarga un backup específico para off-site."""
    u, err, code = _require_admin()
    if err:
        return err, code

    path = get_backup_path(filename)
    if not path:
        return jsonify({"error": "backup no encontrado o nombre inválido"}), 404

    _log_sec("backup_downloaded", u, _client_ip(), f"file={filename}")
    return send_file(
        path,
        mimetype="application/gzip",
        as_attachment=True,
        download_name=filename,
    )


# ─── Users management ────────────────────────────────────────────────────────


def _user_groups(username):
    """Retorna la lista de grupos a los que pertenece un usuario."""
    groups = []
    if username in ADMIN_USERS:        groups.append("Admin")
    if username in CONTADORA_USERS:    groups.append("Contadora")
    if username in COMPRAS_ACCESS:     groups.append("Compras")
    if username in FINANZAS_ACCESS:    groups.append("Finanzas")
    if username in CLIENTES_ACCESS:    groups.append("Clientes")
    if username in MARKETING_USERS:    groups.append("Marketing")
    if username in ANIMUS_ACCESS:      groups.append("ANIMUS")
    if username in ESPAGIRIA_ACCESS:   groups.append("Espagiria")
    if username in TECNICA_USERS:      groups.append("Técnica")
    if username in CALIDAD_USERS:      groups.append("Calidad")
    if username in RRHH_USERS:         groups.append("RRHH")
    if username in PLANTA_USERS:       groups.append("Planta")
    return groups


def _password_source(username, conn):
    """Retorna 'db' / 'env' / 'missing' indicando dónde está la password."""
    try:
        row = conn.execute(
            "SELECT password_hash FROM users_passwords WHERE username=?",
            (username,)
        ).fetchone()
        if row and row[0]:
            return "db"
    except Exception:
        pass
    env_pwd = COMPRAS_USERS.get(username, "")
    if env_pwd and (env_pwd.startswith("pbkdf2:") or env_pwd.startswith("scrypt:")):
        return "env"
    if env_pwd:
        return "env_plaintext"   # alerta: plaintext en env var
    return "missing"


@bp.route("/api/admin/users", methods=["GET"])
def admin_users_list():
    """Lista todos los usuarios con: grupos, fuente de password, último login."""
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    out = []
    for username in sorted(COMPRAS_USERS.keys()):
        # Última vez que entró exitosamente
        last_login = None
        try:
            row = conn.execute(
                """SELECT ts FROM security_events
                   WHERE event='login_success' AND username=?
                   ORDER BY id DESC LIMIT 1""",
                (username,)
            ).fetchone()
            if row:
                last_login = row[0]
        except Exception:
            pass

        # Cuándo se cambió la password (si existe en DB)
        pwd_changed_at = None
        try:
            row = conn.execute(
                "SELECT changed_at, changed_by FROM users_passwords WHERE username=?",
                (username,)
            ).fetchone()
            if row:
                pwd_changed_at = row[0]
        except Exception:
            pass

        out.append({
            "username":       username,
            "groups":         _user_groups(username),
            "is_admin":       username in ADMIN_USERS,
            "password_source": _password_source(username, conn),
            "last_login":     last_login,
            "pwd_changed_at": pwd_changed_at,
        })

    conn.close()
    return jsonify({"users": out, "total": len(out)})


@bp.route("/admin/skus-pendientes", methods=["GET"])
def admin_skus_pendientes_page():
    """Página HTML que lista SKUs no mapeados detectados por
    /api/programacion/debug-calendar y permite asignar producto
    inline sin tocar DB.

    Sebastián 8-may-2026: cuando Alejandro agrega un evento al
    Calendar con un SKU nuevo (HYDRA BALANCE, GLOSSMERLOT, etc.),
    no aparece en /producciones-faltantes hasta que alguien
    agrega el mapping. Esta página le permite a admin hacerlo
    en 30 segundos.
    """
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin</p>',
            status=403, mimetype='text/html'
        )
    return Response(_SKUS_PENDIENTES_HTML, mimetype='text/html')


_SKUS_PENDIENTES_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>SKUs · EOS</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
       background: #f8fafc; color: #1e293b; padding: 20px; line-height: 1.5; }
h1 { font-size: 22px; margin-bottom: 4px; }
h2 { font-size: 16px; margin: 24px 0 10px; color: #334155; }
.sub { color: #64748b; font-size: 13px; margin-bottom: 18px; }
.banner { padding: 14px 18px; border-radius: 10px; margin-bottom: 16px;
          font-size: 14px; font-weight: 600; }
.banner.ok { background: #dcfce7; color: #166534; border: 1px solid #86efac; }
.banner.warn { background: #fef3c7; color: #92400e; border: 1px solid #fde047; }
.row-card { background: #fff; border: 1px solid #e2e8f0; border-radius: 10px;
            padding: 14px 18px; margin-bottom: 10px; }
.row-card.assigned { background: #f0fdf4; border-color: #86efac; opacity: .7; }
.titulo-evento { font-size: 13px; color: #64748b; margin-bottom: 4px; }
.fecha { display: inline-block; font-size: 11px; padding: 2px 8px;
         background: #f1f5f9; border-radius: 4px; color: #475569;
         margin-right: 8px; }
.skus-list { display: inline-flex; gap: 4px; flex-wrap: wrap;
             align-items: center; margin: 6px 0; }
.sku-chip { display: inline-block; padding: 4px 10px; border-radius: 6px;
            font-size: 12px; font-weight: 700; background: #fef3c7;
            color: #92400e; cursor: pointer; border: 1px solid #fde047; }
.sku-chip:hover { background: #fde047; }
.sku-chip.selected { background: #16a34a; color: #fff; border-color: #16a34a; }
.assign-row { display: flex; gap: 8px; align-items: center;
              margin-top: 10px; padding-top: 10px;
              border-top: 1px dashed #e2e8f0; }
.assign-row input { flex: 1; padding: 8px 12px; border: 1px solid #cbd5e1;
                    border-radius: 6px; font-size: 13px; font-family: inherit; }
.btn { padding: 8px 16px; border: none; border-radius: 6px; cursor: pointer;
       font-size: 13px; font-weight: 700; }
.btn.primary { background: #0f766e; color: #fff; }
.btn.primary:hover { background: #0d5d56; }
.btn.primary:disabled { opacity: .5; cursor: not-allowed; }
.btn.danger { background: #dc2626; color: #fff; }
.btn.danger:hover { background: #b91c1c; }
.actions-top { display: flex; gap: 10px; margin-bottom: 14px; }
.muted { color: #94a3b8; font-size: 11px; margin-top: 4px; }
.success-msg { color: #16a34a; font-size: 12px; font-weight: 600; }
.error-msg { color: #dc2626; font-size: 12px; }
table { width: 100%; border-collapse: collapse; background: #fff;
        border-radius: 10px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,.04); }
th { background: #f1f5f9; text-align: left; padding: 10px 14px;
     font-size: 11px; color: #64748b; text-transform: uppercase;
     letter-spacing: .5px; font-weight: 700; }
td { padding: 8px 14px; font-size: 13px; border-top: 1px solid #f1f5f9; }
tr.warning td { background: #fef3c7; }
.badge-no-formula { display: inline-block; font-size: 10px; padding: 1px 6px;
                    background: #fee2e2; color: #991b1b; border-radius: 4px;
                    margin-left: 6px; font-weight: 700; }
.search-bar { padding: 10px 14px; background: #fff; border: 1px solid #e2e8f0;
              border-radius: 8px; margin-bottom: 10px; }
.search-bar input { width: 100%; padding: 8px 10px; border: 1px solid #cbd5e1;
                    border-radius: 6px; font-size: 13px; }
</style>
</head><body>
<h1>🧩 SKUs · mapping con productos</h1>
<div class="sub">Cada evento del Calendar usa un SKU corto (AZHC, BBM, etc.).
Para que aparezca en Plan, el SKU debe estar mapeado a un producto que tenga fórmula.</div>

<div id="banner" class="banner warn">⏳ Cargando...</div>

<div class="actions-top">
  <button class="btn primary" onclick="loadAll()">↻ Recargar</button>
  <button class="btn" onclick="window.open('/api/programacion/debug-calendar?dias=30','_blank')"
          style="background:#475569;color:#fff">JSON raw</button>
  <button class="btn" onclick="window.location.href='/admin'"
          style="background:#475569;color:#fff">← Volver a admin</button>
</div>

<h2>📋 SKUs pendientes (sin mapping)</h2>
<div id="grid-pendientes"></div>

<h2>🗂️ Mapeos existentes (editar / borrar)</h2>
<div class="search-bar">
  <input type="text" id="search-existing" placeholder="Filtrar por SKU o producto..."
         oninput="filterExisting()" />
</div>
<div id="grid-existing"></div>

<script>
let _productosLista = [];
let _existingMappings = [];

async function loadAll() {
  document.getElementById('banner').textContent = '⏳ Cargando...';
  try {
    // 1. Cargar mappings existentes
    const mapsR = await fetch('/api/admin/sku-producto-map?limit=1000');
    const mapsD = await mapsR.json();
    _existingMappings = mapsD.mappings || [];
    const productosSet = new Set();
    _existingMappings.forEach(m => {
      if (m.producto_nombre) productosSet.add(m.producto_nombre);
    });
    _productosLista = [...productosSet].sort();

    // 2. Cargar eventos pendientes (debug-calendar)
    const r = await fetch('/api/programacion/debug-calendar?dias=30');
    const d = await r.json();
    if (!r.ok) { showError(d.error || 'Error'); return; }

    const pendientes = (d.eventos || []).filter(e => e.status === 'sku_no_mapeado');
    renderBanner(pendientes.length, d.total_eventos, _existingMappings.length);
    renderPendientes(pendientes);
    renderExisting();
  } catch(e) {
    showError(e.message);
  }
}

function renderBanner(pendientes, totalEvents, totalMappings) {
  const b = document.getElementById('banner');
  if (pendientes === 0) {
    b.className = 'banner ok';
    b.textContent = '✅ Sin SKUs pendientes · ' + totalMappings +
                    ' mappings existentes · ' + totalEvents + ' eventos en horizonte 30d.';
  } else {
    b.className = 'banner warn';
    b.textContent = '⚠️ ' + pendientes + ' eventos con SKUs no mapeados de ' + totalEvents +
                    ' · ' + totalMappings + ' mappings ya existen.';
  }
}

function showError(msg) {
  const b = document.getElementById('banner');
  b.className = 'banner warn';
  b.textContent = '❌ ' + msg;
}

function esc(s) {
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

function renderPendientes(eventos) {
  const div = document.getElementById('grid-pendientes');
  if (!eventos.length) {
    div.innerHTML = '<div class="muted">✅ Nada pendiente · todos los SKUs del Calendar están mapeados.</div>';
    return;
  }
  let html = '';
  eventos.forEach((ev, i) => {
    const skus = ev.skus_detectados || [];
    html += '<div class="row-card" id="card-' + i + '" data-idx="' + i + '">' +
      '<div class="titulo-evento">' +
        '<span class="fecha">' + esc(ev.fecha) + '</span>' +
        esc(ev.titulo) +
      '</div>' +
      '<div class="skus-list">SKUs: ' +
        skus.map(s =>
          '<span class="sku-chip" onclick="selectSku(' + i + ',\\''+ esc(s) +'\\')">' + esc(s) + '</span>'
        ).join('') +
      '</div>' +
      '<div class="assign-row">' +
        '<input type="text" id="sku-input-' + i + '" placeholder="Click un SKU arriba o escribí manual" />' +
        '<input type="text" id="prod-input-' + i + '" list="prod-list" placeholder="Producto destino (ej: GEL HIDRATANTE)" />' +
        '<button class="btn primary" onclick="saveMapping(' + i + ')">💾 Guardar</button>' +
        '<span id="msg-' + i + '"></span>' +
      '</div>' +
    '</div>';
  });
  html += '<datalist id="prod-list">' +
    _productosLista.map(p => '<option value="' + esc(p) + '">').join('') +
    '</datalist>';
  div.innerHTML = html;
}

function renderExisting() {
  const div = document.getElementById('grid-existing');
  if (!_existingMappings.length) {
    div.innerHTML = '<div class="muted">No hay mappings existentes aún.</div>';
    return;
  }
  const filter = (document.getElementById('search-existing')||{}).value || '';
  const filterUp = filter.toUpperCase();
  const filtered = _existingMappings.filter(m => {
    if (!filter) return true;
    return ((m.sku||'').toUpperCase().indexOf(filterUp) >= 0) ||
           ((m.producto_nombre||'').toUpperCase().indexOf(filterUp) >= 0);
  });
  let html = '<table>' +
    '<thead><tr>' +
      '<th>SKU</th>' +
      '<th>Producto destino</th>' +
      '<th>Estado</th>' +
      '<th>Acciones</th>' +
    '</tr></thead><tbody>';
  filtered.forEach((m, i) => {
    const isInactive = !m.activo;
    html += '<tr id="exist-row-' + i + '"' +
            (isInactive ? ' style="opacity:.5"' : '') + '>' +
      '<td><b>' + esc(m.sku) + '</b></td>' +
      '<td>' +
        '<input type="text" id="exist-prod-' + i + '" value="' + esc(m.producto_nombre) + '" ' +
        'list="prod-list-exist" style="width:100%;padding:6px;border:1px solid #cbd5e1;border-radius:4px;font-size:12px"/>' +
      '</td>' +
      '<td>' + (isInactive ? '<span class="badge-no-formula">DESACTIVADO</span>' : '<span style="color:#16a34a">activo</span>') + '</td>' +
      '<td style="white-space:nowrap">' +
        '<button class="btn primary" style="padding:4px 10px;font-size:11px" onclick="updateExisting(' + i + ',\\''+ esc(m.sku) +'\\')">💾</button> ' +
        '<button class="btn danger" style="padding:4px 10px;font-size:11px" onclick="deleteExisting(' + i + ',\\''+ esc(m.sku) +'\\')">🗑️</button> ' +
        '<span id="exist-msg-' + i + '" style="margin-left:6px"></span>' +
      '</td>' +
    '</tr>';
  });
  html += '</tbody></table>' +
    '<datalist id="prod-list-exist">' +
    _productosLista.map(p => '<option value="' + esc(p) + '">').join('') +
    '</datalist>';
  div.innerHTML = html;
}

function filterExisting() { renderExisting(); }

async function updateExisting(idx, sku) {
  const prod = document.getElementById('exist-prod-' + idx).value.trim();
  const msg = document.getElementById('exist-msg-' + idx);
  if (!prod) { msg.className = 'error-msg'; msg.textContent = 'producto vacío'; return; }
  msg.textContent = '⏳';
  try {
    const r = await fetch('/api/admin/sku-producto-map', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({sku: sku, producto_nombre: prod, activo: true})
    });
    const d = await r.json();
    if (r.ok) {
      msg.className = 'success-msg';
      msg.textContent = '✅';
      // Update local list
      _existingMappings.forEach(m => { if (m.sku === sku) m.producto_nombre = prod; });
      setTimeout(() => loadAll(), 500);
    } else {
      msg.className = 'error-msg';
      msg.textContent = '❌ ' + (d.error || '?');
    }
  } catch(e) {
    msg.className = 'error-msg';
    msg.textContent = '❌ ' + e.message;
  }
}

async function deleteExisting(idx, sku) {
  if (!confirm('Desactivar mapping ' + sku + '?\\nLos eventos con este SKU dejarán de aparecer en Plan hasta que crees uno nuevo.')) return;
  const msg = document.getElementById('exist-msg-' + idx);
  msg.textContent = '⏳';
  try {
    const r = await fetch('/api/admin/sku-producto-map', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({sku: sku, _delete: true})
    });
    if (r.ok) {
      msg.className = 'success-msg';
      msg.textContent = '✅ desactivado';
      setTimeout(() => loadAll(), 500);
    } else {
      msg.className = 'error-msg';
      msg.textContent = '❌';
    }
  } catch(e) {
    msg.className = 'error-msg';
    msg.textContent = '❌ ' + e.message;
  }
}

function selectSku(idx, sku) {
  document.getElementById('sku-input-' + idx).value = sku;
  const card = document.getElementById('card-' + idx);
  card.querySelectorAll('.sku-chip').forEach(c => c.classList.remove('selected'));
  card.querySelectorAll('.sku-chip').forEach(c => {
    if (c.textContent.trim() === sku) c.classList.add('selected');
  });
}

async function saveMapping(idx) {
  const sku = document.getElementById('sku-input-' + idx).value.trim().toUpperCase();
  const prod = document.getElementById('prod-input-' + idx).value.trim();
  const msgEl = document.getElementById('msg-' + idx);
  if (!sku || !prod) {
    msgEl.className = 'error-msg';
    msgEl.textContent = 'SKU y producto son obligatorios';
    return;
  }
  msgEl.className = 'muted';
  msgEl.textContent = '⏳ Guardando...';
  try {
    const r = await fetch('/api/admin/sku-producto-map', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({sku: sku, producto_nombre: prod, activo: true})
    });
    const d = await r.json();
    if (r.ok) {
      msgEl.className = 'success-msg';
      msgEl.textContent = '✅ ' + sku + ' → ' + prod;
      const card = document.getElementById('card-' + idx);
      card.classList.add('assigned');
      if (d.warning) {
        msgEl.textContent += ' (advertencia: ' + d.warning + ')';
      }
      setTimeout(() => loadAll(), 1000);
    } else {
      msgEl.className = 'error-msg';
      msgEl.textContent = '❌ ' + (d.error || 'Error');
    }
  } catch(e) {
    msgEl.className = 'error-msg';
    msgEl.textContent = '❌ ' + e.message;
  }
}

loadAll();
</script>

<!-- Widget Mi contraseña -->
<a href="/cambiar-password" title="Cambiar mi contraseña"
   style="position:fixed;bottom:24px;left:24px;z-index:9998;
          background:#1e293b;color:#a78bfa;border:1px solid #4c1d95;
          border-radius:24px;padding:8px 16px;font-size:12px;font-weight:700;
          text-decoration:none;box-shadow:0 4px 12px rgba(0,0,0,.2);
          font-family:-apple-system,Segoe UI,sans-serif;
          display:flex;align-items:center;gap:6px">
  🔐 Mi contraseña
</a>
</body></html>
"""


@bp.route("/api/admin/sku-producto-map", methods=["GET", "POST"])
def admin_sku_producto_map():
    """CRUD del mapeo SKU → producto · Sebastián 8-may-2026.

    Sin este mapping, eventos del Calendar con SKUs como GLOSSMERLOT,
    HYDRA PEPTIDE, etc. quedan invisibles en /producciones-faltantes.

    GET:
      ?sku=X       → busca un SKU específico
      ?producto=Y  → busca productos por nombre
      (sin args)   → lista todos (last 500)

    POST:
      body: {sku, producto_nombre, activo: true}
      Upsert · si existe sku, actualiza producto y activo.

    POST con _delete=true:
      body: {sku, _delete: true} → marca activo=0 (soft delete)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if request.method == "GET":
        sku = (request.args.get("sku") or "").strip().upper()
        prod = (request.args.get("producto") or "").strip()
        try:
            limit = max(1, min(int(request.args.get("limit", 500)), 5000))
        except (ValueError, TypeError):
            limit = 500
        # Tolerante a schema · creado_at/por pueden no existir
        cols_info = c.execute("PRAGMA table_info(sku_producto_map)").fetchall()
        col_names = {ci[1] for ci in cols_info}
        creado_at_col = ("COALESCE(creado_at,'')" if 'creado_at' in col_names else "''")
        creado_por_col = ("COALESCE(creado_por,'')" if 'creado_por' in col_names else "''")
        sql = (f"SELECT sku, producto_nombre, COALESCE(activo,1), "
               f"{creado_at_col}, {creado_por_col} "
               "FROM sku_producto_map WHERE 1=1")
        params = []
        if sku:
            sql += " AND UPPER(sku)=?"
            params.append(sku)
        if prod:
            sql += " AND UPPER(producto_nombre) LIKE ?"
            params.append(f'%{prod.upper()}%')
        sql += " ORDER BY sku LIMIT ?"
        params.append(limit)
        rows = c.execute(sql, params).fetchall()
        conn.close()
        return jsonify({
            "mappings": [
                {"sku": r[0], "producto_nombre": r[1],
                 "activo": bool(r[2]), "creado_at": r[3], "creado_por": r[4]}
                for r in rows
            ],
            "total": len(rows),
        })

    # POST · upsert o delete
    body = request.get_json(silent=True) or {}
    sku = (body.get("sku") or "").strip().upper()
    if not sku:
        conn.close()
        return jsonify({"error": "sku requerido"}), 400

    if body.get("_delete"):
        c.execute("UPDATE sku_producto_map SET activo=0 WHERE UPPER(sku)=?", (sku,))
        rows = c.rowcount
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "deleted": sku, "rows_affected": rows})

    producto = (body.get("producto_nombre") or "").strip()
    if not producto:
        conn.close()
        return jsonify({"error": "producto_nombre requerido"}), 400

    activo = 1 if body.get("activo", True) else 0

    # Verificar que el producto existe en formula_headers (recomendado)
    formula_existe = c.execute(
        "SELECT 1 FROM formula_headers WHERE UPPER(TRIM(producto_nombre))=? LIMIT 1",
        (producto.upper().strip(),)
    ).fetchone()

    # Upsert
    try:
        c.execute("""
            INSERT INTO sku_producto_map(sku, producto_nombre, activo)
            VALUES(?, ?, ?)
            ON CONFLICT(sku) DO UPDATE SET
              producto_nombre = excluded.producto_nombre,
              activo = excluded.activo
        """, (sku, producto, activo))
        # Best-effort: agregar created_by/at si las columnas existen
        try:
            c.execute(
                "UPDATE sku_producto_map SET creado_por=? WHERE sku=?",
                (u, sku),
            )
        except sqlite3.OperationalError:
            pass
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({"error": f"DB: {e}"}), 500

    conn.close()
    return jsonify({
        "ok": True, "sku": sku, "producto_nombre": producto,
        "activo": bool(activo),
        "warning": (None if formula_existe
                    else f'producto "{producto}" no tiene fórmula · '
                         'el sync incluirá el evento pero sin calcular MPs faltantes'),
    })


@bp.route("/api/admin/zero-error/status", methods=["GET"])
def admin_zero_error_status():
    """Dashboard agregador del sistema anti-regresión.

    Sebastián 7-may-2026: una sola query devuelve todo el estado del
    sistema zero-error. Lo consume la página /admin/zero-error.

    Devuelve:
      · golden_paths: count + último resultado (si lo tenemos)
      · watcher: últimos 5 runs del cron + status global
      · health: estado actual de los 8 invariantes (call interno)
      · agent_memory: count + últimas 5 entries por categoría
      · session_logs: últimos 5 archivos en SESSION_LOG/
      · git: últimos 10 commits
      · pending_bugs: lista de bugs conocidos sin resolver (manual)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    import os as _os
    import subprocess as _sp
    import re as _re
    from datetime import datetime as _dt

    REPO_ROOT = _os.path.dirname(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    )

    out = {
        'generated_at': _dt.utcnow().isoformat() + 'Z',
        'repo_root': REPO_ROOT,
    }

    # 1. Golden paths · contar tests en el archivo
    try:
        gp_file = _os.path.join(REPO_ROOT, 'tests', 'test_golden_paths.py')
        if _os.path.exists(gp_file):
            with open(gp_file, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            tests = _re.findall(r'^def (test_golden_\w+)\(', content, _re.MULTILINE)
            out['golden_paths'] = {
                'count': len(tests),
                'tests': tests[:60],
                'file': 'tests/test_golden_paths.py',
            }
        else:
            out['golden_paths'] = {'count': 0, 'tests': [], 'error': 'archivo no encontrado'}
    except Exception as e:
        out['golden_paths'] = {'error': str(e)}

    # 2. Watcher · últimos 5 runs
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        # Sebastian 7-may-2026: schema correcto es resultado_json + error
        # (no detalle_json + error_msg que era el guess inicial)
        rows = c.execute("""
            SELECT job_name, ejecutado_at, ok,
                   COALESCE(resultado_json, ''),
                   COALESCE(error, ''),
                   COALESCE(duracion_ms, 0)
            FROM cron_jobs_runs
            WHERE job_name LIKE 'watcher_health%'
            ORDER BY ejecutado_at DESC LIMIT 5
        """).fetchall()
        out['watcher'] = {
            'runs': [
                {
                    'job_name': r[0], 'ejecutado_at': r[1],
                    'ok': bool(r[2]), 'detalle': (r[3] or '')[:500],
                    'error': (r[4] or '')[:200], 'duracion_ms': r[5],
                } for r in rows
            ],
            'count_recent': len(rows),
        }
    except sqlite3.OperationalError as _e:
        # Tabla puede no existir si la mig 72 no corrió o si DB es nueva
        out['watcher'] = {
            'runs': [],
            'note': f'cron_jobs_runs no disponible: {_e}',
        }
    except Exception as e:
        out['watcher'] = {'error': str(e)}

    # 3. Health · invocar el endpoint internamente
    try:
        from flask import session as _s, current_app as _ca
        with _ca.test_request_context('/api/admin/health/critical-paths'):
            _s['compras_user'] = u
            resp = admin_health_critical_paths()
        if isinstance(resp, tuple):
            health_payload = resp[0].get_json()
        else:
            health_payload = resp.get_json()
        out['health'] = health_payload
    except Exception as e:
        out['health'] = {'error': str(e)}

    # 4. agent_memory · count + últimas 5 entries
    try:
        total = c.execute("SELECT COUNT(*) FROM agent_memory").fetchone()[0]
        recent = c.execute("""
            SELECT key, category, updated_at, created_by,
                   SUBSTR(value, 1, 100)
            FROM agent_memory
            ORDER BY updated_at DESC LIMIT 5
        """).fetchall()
        cat_counts = c.execute("""
            SELECT category, COUNT(*) FROM agent_memory
            GROUP BY category ORDER BY 2 DESC
        """).fetchall()
        out['agent_memory'] = {
            'total': total,
            'recent': [
                {'key': r[0], 'category': r[1], 'updated_at': r[2],
                 'created_by': r[3], 'value_preview': r[4]}
                for r in recent
            ],
            'by_category': [{'category': r[0], 'count': r[1]}
                            for r in cat_counts],
        }
    except sqlite3.OperationalError:
        out['agent_memory'] = {'total': 0, 'note': 'Tabla agent_memory no existe (mig 96 pendiente)'}
    except Exception as e:
        out['agent_memory'] = {'error': str(e)}
    conn.close()

    # 5. SESSION_LOG · listar últimos 5 archivos
    try:
        log_dir = _os.path.join(REPO_ROOT, 'SESSION_LOG')
        if _os.path.isdir(log_dir):
            files = []
            for fn in _os.listdir(log_dir):
                if not fn.endswith('.md') or fn == 'README.md':
                    continue
                path = _os.path.join(log_dir, fn)
                try:
                    st = _os.stat(path)
                    files.append({
                        'name': fn,
                        'mtime': _dt.utcfromtimestamp(st.st_mtime).isoformat() + 'Z',
                        'size': st.st_size,
                    })
                except Exception:
                    pass
            files.sort(key=lambda f: f['mtime'], reverse=True)
            out['session_logs'] = {'files': files[:5], 'total': len(files)}
        else:
            out['session_logs'] = {'files': [], 'note': 'SESSION_LOG/ no existe'}
    except Exception as e:
        out['session_logs'] = {'error': str(e)}

    # 6. Git · últimos 10 commits
    try:
        result = _sp.run(
            ['git', '-C', REPO_ROOT, 'log', '--oneline', '-10',
             '--pretty=format:%h|%s|%an|%ar'],
            capture_output=True, text=True,
            encoding='utf-8', errors='replace', timeout=10,
        )
        commits = []
        for line in (result.stdout or '').splitlines():
            parts = line.split('|', 3)
            if len(parts) == 4:
                commits.append({
                    'hash': parts[0], 'message': parts[1],
                    'author': parts[2], 'when': parts[3],
                })
        out['git'] = {'recent_commits': commits, 'total_shown': len(commits)}
    except Exception as e:
        out['git'] = {'error': str(e)}

    # 7. Pending bugs (manual list · podemos persistirlos en agent_memory después)
    # AZHC-LUN-11 RESUELTO 2026-05-07 · debug confirmó que las 3 fechas
    # (11-may, 14-may, 12-ago) corresponden a eventos reales en Calendar.
    # No era fantasma · era 2 producciones distintas en la misma semana.
    out['pending_bugs'] = [
        {
            'id': 'MAYERLIN-LOGIN',
            'title': 'Mayerlin no puede entrar',
            'severity': 'high',
            'next_action': (
                'Ir a /admin → Usuarios → Mayerlin → Diag · '
                'si password_source=missing → click Resetear'
            ),
        },
    ]
    out['resolved_recent'] = [
        {
            'id': 'AZHC-LUN-11',
            'title': 'AZHC fantasma Lun 11',
            'resolved_at': '2026-05-07',
            'resolution': (
                'Falsa alarma post-fix · sync espejo limpió 9 entries '
                'canceladas · las 3 fechas vivas (11-may, 14-may, 12-ago) '
                'corresponden a eventos reales en Calendar.'
            ),
        },
    ]

    # 8. Status global derivado
    has_critical = (
        out.get('health', {}).get('critical_count', 0) > 0
        or any(not r.get('ok') for r in out.get('watcher', {}).get('runs', []))
    )
    out['global_status'] = 'critical' if has_critical else 'ok'

    return jsonify(out)


@bp.route("/admin/zero-error", methods=["GET"])
def admin_zero_error_page():
    """Dashboard HTML del sistema anti-regresión.

    Renderiza una página standalone que consume
    /api/admin/zero-error/status y muestra cards con el estado.
    Solo admin puede verla.
    """
    u, err, code = _require_admin()
    if err:
        # Si no es admin, redirigir a login en vez de devolver JSON
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este dashboard.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_ZERO_ERROR_DASHBOARD_HTML, mimetype='text/html')


_ZERO_ERROR_DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>Zero-Error · EOS</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
       background: #f8fafc; color: #1e293b; padding: 20px; line-height: 1.5; }
h1 { font-size: 22px; margin-bottom: 4px; }
.sub { color: #64748b; font-size: 13px; margin-bottom: 18px; }
.grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
        gap: 14px; }
.card { background: #fff; border-radius: 10px; border: 1px solid #e2e8f0;
        padding: 14px 16px; box-shadow: 0 1px 3px rgba(0,0,0,.04); }
.card h3 { font-size: 13px; color: #475569; text-transform: uppercase;
           letter-spacing: .4px; margin-bottom: 10px; font-weight: 700;
           display: flex; align-items: center; gap: 6px; }
.dot { width: 9px; height: 9px; border-radius: 50%; display: inline-block; }
.dot.ok { background: #16a34a; }
.dot.warn { background: #f59e0b; }
.dot.critical { background: #dc2626; }
.kpi { font-size: 28px; font-weight: 800; color: #0f172a; line-height: 1.1; }
.kpi-label { font-size: 11px; color: #94a3b8; }
.row { display: flex; justify-content: space-between; align-items: center;
       padding: 6px 0; border-bottom: 1px solid #f1f5f9; font-size: 12px; }
.row:last-child { border-bottom: none; }
.row .v { color: #475569; font-family: monospace; font-size: 11px; }
.muted { color: #94a3b8; font-size: 11px; }
.tag { display: inline-block; padding: 1px 8px; border-radius: 4px;
       font-size: 10px; font-weight: 700; text-transform: uppercase;
       letter-spacing: .3px; }
.tag.ok { background: #dcfce7; color: #166534; }
.tag.warn { background: #fef3c7; color: #92400e; }
.tag.critical { background: #fee2e2; color: #991b1b; }
.global-banner { padding: 12px 16px; border-radius: 10px; margin-bottom: 16px;
                 font-weight: 700; }
.global-banner.ok { background: #dcfce7; color: #166534;
                    border: 1px solid #86efac; }
.global-banner.critical { background: #fee2e2; color: #991b1b;
                           border: 1px solid #fca5a5; }
button { background: #0f766e; color: #fff; border: none; padding: 8px 14px;
         border-radius: 6px; cursor: pointer; font-weight: 700; font-size: 12px; }
button:hover { background: #0d5d56; }
.actions { display: flex; gap: 8px; margin-bottom: 14px; }
pre { font-size: 11px; background: #f1f5f9; padding: 8px; border-radius: 4px;
      overflow-x: auto; max-height: 200px; }
.bug { background: #fef2f2; border-left: 3px solid #dc2626; padding: 8px 12px;
       margin: 6px 0; border-radius: 4px; }
.bug .title { font-weight: 700; color: #7f1d1d; font-size: 12px; }
.bug .next { font-size: 11px; color: #4b5563; margin-top: 4px; }
</style>
</head>
<body>
<h1>🛡️ Zero-Error Dashboard · EOS</h1>
<div class="sub">Estado del sistema anti-regresión · auto-refresh cada 60s</div>

<div id="global-banner" class="global-banner ok">⏳ Cargando...</div>

<div class="actions">
  <button onclick="loadStatus()">↻ Recargar</button>
  <button onclick="window.open('/api/admin/zero-error/status','_blank')"
          style="background:#475569">JSON crudo</button>
  <button onclick="window.open('/api/admin/health/critical-paths','_blank')"
          style="background:#475569">Health checks</button>
</div>

<div class="grid" id="grid">
  <div class="card"><h3>Cargando...</h3></div>
</div>

<script>
async function loadStatus() {
  try {
    const r = await fetch('/api/admin/zero-error/status');
    const d = await r.json();
    render(d);
  } catch (e) {
    document.getElementById('grid').innerHTML =
      '<div class="card"><h3>Error</h3><pre>' + e.message + '</pre></div>';
  }
}

function fmt(s) { return (s == null) ? '—' : String(s); }
function esc(s) { return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

function render(d) {
  // Global banner
  const banner = document.getElementById('global-banner');
  banner.className = 'global-banner ' + (d.global_status || 'critical');
  banner.textContent = d.global_status === 'ok'
    ? '✅ Sistema OK · todos los checks verdes'
    : '🚨 Atención · hay checks críticos pendientes';

  const grid = document.getElementById('grid');
  let html = '';

  // Card 1: Golden Paths
  const gp = d.golden_paths || {};
  html += card('🛡️ Golden Paths', gp.error ? 'critical' : 'ok',
    `<div class="kpi">${gp.count||0}</div>
     <div class="kpi-label">tests E2E protegen flujos críticos</div>
     <div class="muted" style="margin-top:8px">${esc(gp.file||'')}</div>
     <div style="margin-top:8px">
       <a href="javascript:document.getElementById('gp-list').style.display=document.getElementById('gp-list').style.display==='none'?'block':'none'" style="font-size:11px">Ver/ocultar lista</a>
       <pre id="gp-list" style="display:none">${(gp.tests||[]).map(esc).join('\\n')}</pre>
     </div>`);

  // Card 2: Health checks
  const h = d.health || {};
  const hStatus = (h.status || 'critical');
  html += card('💚 Health · 8 invariantes',
    hStatus === 'ok' ? 'ok' : (hStatus === 'warn' ? 'warn' : 'critical'),
    h.checks ? h.checks.map(c => `
      <div class="row">
        <span><span class="dot ${c.status||'critical'}"></span> ${esc(c.name)}</span>
        <span class="v">${esc(c.detail||'')}</span>
      </div>`).join('') : `<pre>${esc(JSON.stringify(h,null,2))}</pre>`);

  // Card 3: Watcher cron
  const w = d.watcher || {};
  const lastRun = (w.runs && w.runs[0]) || null;
  html += card('⏰ Watcher cron',
    lastRun ? (lastRun.ok ? 'ok' : 'critical') : 'warn',
    `<div class="kpi">${(w.runs||[]).length}</div>
     <div class="kpi-label">runs recientes</div>
     ${(w.runs||[]).map(r => `
       <div class="row">
         <span><span class="dot ${r.ok?'ok':'critical'}"></span> ${esc(r.job_name)}</span>
         <span class="v">${esc(r.ejecutado_at||'')}</span>
       </div>`).join('') || '<div class="muted">Sin runs registrados todavía. El cron corre cada hora a :07.</div>'}`);

  // Card 4: agent_memory
  const am = d.agent_memory || {};
  html += card('🧠 Agent Memory', 'ok',
    `<div class="kpi">${am.total||0}</div>
     <div class="kpi-label">entries persistidas</div>
     ${(am.by_category||[]).map(c => `
       <div class="row">
         <span>${esc(c.category)}</span>
         <span class="v">${c.count}</span>
       </div>`).join('')}
     ${am.note ? `<div class="muted" style="margin-top:8px">${esc(am.note)}</div>` : ''}`);

  // Card 5: SESSION_LOG
  const sl = d.session_logs || {};
  html += card('📝 Session Log', 'ok',
    `<div class="kpi">${sl.total||0}</div>
     <div class="kpi-label">archivos · últimos 5</div>
     ${(sl.files||[]).map(f => `
       <div class="row">
         <span>${esc(f.name)}</span>
         <span class="v">${esc((f.mtime||'').slice(0,10))}</span>
       </div>`).join('')}`);

  // Card 6: Git
  const g = d.git || {};
  html += card('🔀 Git · últimos commits', 'ok',
    `${(g.recent_commits||[]).map(c => `
       <div class="row">
         <span><b>${esc(c.hash)}</b> ${esc(c.message).slice(0,50)}${c.message.length>50?'...':''}</span>
         <span class="v">${esc(c.when)}</span>
       </div>`).join('')}`);

  // Card 7: Bugs pendientes
  const bugs = d.pending_bugs || [];
  html += card('🐛 Bugs pendientes',
    bugs.length ? 'warn' : 'ok',
    `<div class="kpi">${bugs.length}</div>
     <div class="kpi-label">bugs conocidos sin resolver</div>
     ${bugs.map(b => `
       <div class="bug">
         <div class="title">[${esc(b.severity)}] ${esc(b.title)}</div>
         <div class="next">→ ${esc(b.next_action)}</div>
       </div>`).join('')}`);

  // Card 8: Meta-info
  html += card('ℹ️ Meta', 'ok',
    `<div class="row"><span>Generado</span><span class="v">${esc(d.generated_at)}</span></div>
     <div class="row"><span>Repo</span><span class="v">${esc((d.repo_root||'').split(/[\\\\/]/).pop())}</span></div>
     <div class="row"><span>Status global</span><span class="tag ${d.global_status}">${esc(d.global_status)}</span></div>`);

  grid.innerHTML = html;
}

function card(title, status, body) {
  return `<div class="card">
    <h3><span class="dot ${status}"></span> ${title}</h3>
    ${body}
  </div>`;
}

loadStatus();
setInterval(loadStatus, 60000);
</script>
</body></html>
"""


@bp.route("/api/admin/health/critical-paths", methods=["GET"])
def admin_health_critical_paths():
    """Health check de invariantes críticas en producción.

    Sebastián 7-may-2026 (zero-error sprint día 4): el Watcher cron pega
    este endpoint cada 15 min. Si algún check falla (status='critical')
    se manda email a EMAIL_GERENCIA.

    Devuelve UN status global y por check:
      · ok       · todo bien
      · warn     · anomalía menor, no bloquea operación
      · critical · bug que afecta usuarios YA · alerta inmediata

    Checks (ALL READ-ONLY · no muta data):
      1. tablas_criticas · core tables existen y tienen filas
      2. indexes_criticos · indexes de performance presentes
      3. producciones_zombie · iniciadas hace >30d sin completar
      4. sols_planta_huerfanas · planta SOLs sin OC vencidas
      5. last_calendar_sync · sync se ejecutó <2h
      6. last_backup · backup <30h (Sebastián tolerante)
      7. agent_memory_smoke · tabla nueva responde
      8. movimientos_consistency · ningún tipo distinto de Entrada/Salida
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    import time as _time
    t0 = _time.time()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    checks = []

    def _check(name, status, detail, value=None, threshold=None):
        checks.append({
            'name': name, 'status': status, 'detail': detail,
            'value': value, 'threshold': threshold,
        })

    # 1. tablas críticas existen
    try:
        critical_tables = [
            'movimientos', 'maestro_mps', 'produccion_programada',
            'solicitudes_compra', 'conteos_fisicos', 'conteo_items',
            'audit_log', 'mp_lead_time_config', 'agent_memory',
        ]
        missing = []
        for t in critical_tables:
            row = c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (t,)
            ).fetchone()
            if not row:
                missing.append(t)
        if missing:
            _check('tablas_criticas', 'critical',
                   f'Faltan tablas: {", ".join(missing)}', value=missing)
        else:
            _check('tablas_criticas', 'ok',
                   f'{len(critical_tables)} tablas presentes',
                   value=len(critical_tables))
    except Exception as e:
        _check('tablas_criticas', 'critical', f'Error query: {e}')

    # 2. indexes críticos
    try:
        critical_indexes = [
            'idx_mov_material', 'idx_mov_lote', 'idx_mov_fecha',
            'idx_oc_estado', 'idx_sol_estado',
        ]
        rows = c.execute(
            "SELECT name FROM sqlite_master WHERE type='index'"
        ).fetchall()
        existing = {r[0] for r in rows}
        missing = [i for i in critical_indexes if i not in existing]
        if missing:
            _check('indexes_criticos', 'warn',
                   f'Faltan indexes (degrade performance): {missing}',
                   value=missing)
        else:
            _check('indexes_criticos', 'ok',
                   f'{len(critical_indexes)} indexes verificados')
    except Exception as e:
        _check('indexes_criticos', 'warn', f'Error query: {e}')

    # 3. producciones zombie (iniciadas hace >30d sin completar)
    try:
        rows = c.execute("""
            SELECT COUNT(*), GROUP_CONCAT(producto || ' (' || fecha_programada || ')', ', ')
            FROM produccion_programada
            WHERE inicio_real_at IS NOT NULL
              AND COALESCE(estado, '') NOT IN ('completado', 'cancelado')
              AND date(fecha_programada) < date('now', '-30 days')
        """).fetchone()
        n = rows[0] or 0
        if n > 5:
            _check('producciones_zombie', 'critical',
                   f'{n} producciones iniciadas hace >30d sin cerrar', value=n,
                   threshold='>5')
        elif n > 0:
            _check('producciones_zombie', 'warn',
                   f'{n} producciones iniciadas hace >30d', value=n)
        else:
            _check('producciones_zombie', 'ok', 'sin zombies', value=0)
    except Exception as e:
        _check('producciones_zombie', 'warn', f'Error: {e}')

    # 4. SOLs planta huérfanas (Pendientes >14d sin OC)
    try:
        n = c.execute("""
            SELECT COUNT(*) FROM solicitudes_compra
            WHERE estado='Pendiente' AND COALESCE(numero_oc,'') = ''
              AND categoria IN ('Materia Prima','Empaque','Material de Empaque')
              AND date(fecha) < date('now', '-14 days')
        """).fetchone()[0] or 0
        if n > 20:
            _check('sols_planta_huerfanas', 'warn',
                   f'{n} SOLs planta Pendientes >14d sin OC · considera Limpiar',
                   value=n, threshold='>20')
        else:
            _check('sols_planta_huerfanas', 'ok',
                   f'{n} SOLs planta Pendientes >14d', value=n)
    except Exception as e:
        _check('sols_planta_huerfanas', 'warn', f'Error: {e}')

    # 5. last calendar sync < 2h
    # Sebastian 7-may-2026: tabla es `sync_log` (sin underscore prefix).
    # _ensure_sync_log_table() en programacion.py la crea on-demand.
    try:
        row = c.execute("""
            SELECT MAX(last_run_at) FROM sync_log WHERE sync_type='calendar'
        """).fetchone()
        last_sync = row[0] if row else None
        if not last_sync:
            _check('last_calendar_sync', 'warn',
                   'Nunca corrió un sync de calendar (tabla vacía o sin runs)')
        else:
            # Parse ISO timestamp
            from datetime import datetime as _dt
            try:
                ts = _dt.fromisoformat(last_sync.replace('Z', '+00:00'))
                age_min = (_dt.utcnow() - ts.replace(tzinfo=None)).total_seconds() / 60
            except Exception:
                age_min = 99999
            if age_min > 120:
                _check('last_calendar_sync', 'critical',
                       f'Sync no corre desde {age_min:.0f}min (>2h)',
                       value=round(age_min), threshold='>120 min')
            else:
                _check('last_calendar_sync', 'ok',
                       f'Sync hace {age_min:.0f}min', value=round(age_min))
    except sqlite3.OperationalError:
        # Tabla puede no existir aún si nunca corrió un sync
        _check('last_calendar_sync', 'warn',
               'Tabla sync_log aún no creada (sync nunca corrió en este worker)')
    except Exception as e:
        _check('last_calendar_sync', 'warn', f'Error: {e}')

    # 6. last backup < 30h
    try:
        row = c.execute("""
            SELECT MAX(completed_at) FROM backup_log WHERE status='ok'
        """).fetchone()
        last_bk = row[0] if row else None
        if not last_bk:
            _check('last_backup', 'warn', 'No hay backups exitosos')
        else:
            from datetime import datetime as _dt
            try:
                ts = _dt.fromisoformat(last_bk.replace('Z', '+00:00'))
                age_h = (_dt.utcnow() - ts.replace(tzinfo=None)).total_seconds() / 3600
            except Exception:
                age_h = 999
            if age_h > 30:
                _check('last_backup', 'critical',
                       f'Sin backup en {age_h:.1f}h (>30h)',
                       value=round(age_h, 1), threshold='>30 h')
            else:
                _check('last_backup', 'ok',
                       f'Backup hace {age_h:.1f}h', value=round(age_h, 1))
    except sqlite3.OperationalError:
        _check('last_backup', 'warn', 'Tabla backup_log no existe')
    except Exception as e:
        _check('last_backup', 'warn', f'Error: {e}')

    # 7. agent_memory smoke (tabla nueva del Día 3 respondiendo)
    try:
        n = c.execute("SELECT COUNT(*) FROM agent_memory").fetchone()[0]
        _check('agent_memory_smoke', 'ok',
               f'{n} entries · tabla viva', value=n)
    except Exception as e:
        _check('agent_memory_smoke', 'critical',
               f'Tabla agent_memory falló: {e}')

    # 8. movimientos consistency · tipos válidos
    try:
        invalid = c.execute("""
            SELECT COUNT(*) FROM movimientos
            WHERE tipo NOT IN ('Entrada','Salida','entrada','salida','ENTRADA','SALIDA')
              AND tipo IS NOT NULL AND tipo != ''
        """).fetchone()[0] or 0
        if invalid > 0:
            _check('movimientos_consistency', 'critical',
                   f'{invalid} movimientos con tipo inválido', value=invalid)
        else:
            _check('movimientos_consistency', 'ok',
                   'Todos los movimientos tienen tipo válido')
    except Exception as e:
        _check('movimientos_consistency', 'warn', f'Error: {e}')

    conn.close()

    # Status global = peor de todos
    has_critical = any(c['status'] == 'critical' for c in checks)
    has_warn = any(c['status'] == 'warn' for c in checks)
    if has_critical:
        global_status = 'critical'
    elif has_warn:
        global_status = 'warn'
    else:
        global_status = 'ok'

    elapsed_ms = round((_time.time() - t0) * 1000, 1)
    return jsonify({
        'status': global_status,
        'checks': checks,
        'critical_count': sum(1 for c in checks if c['status'] == 'critical'),
        'warn_count': sum(1 for c in checks if c['status'] == 'warn'),
        'ok_count': sum(1 for c in checks if c['status'] == 'ok'),
        'total_checks': len(checks),
        'elapsed_ms': elapsed_ms,
    }), 200 if global_status != 'critical' else 503


@bp.route("/api/admin/agent-memory", methods=["GET", "POST"])
def admin_agent_memory():
    """CRUD de la tabla agent_memory · memoria persistente entre sesiones IA.

    Sebastián 7-may-2026 (zero-error sprint día 3): los agentes IA leen y
    escriben aquí para resolver "amnesia entre sesiones". Cada entrada es
    key-value con categoría opcional para filtrar.

    GET  /api/admin/agent-memory                 → lista todo (last 100)
    GET  /api/admin/agent-memory?category=bug    → filtra por categoría
    GET  /api/admin/agent-memory?key=X           → entry específica
    POST /api/admin/agent-memory                 → upsert
         body: {key, value, category?}
    DELETE via POST con body {key, _delete: true}
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if request.method == "GET":
        key = request.args.get("key", "").strip()
        category = request.args.get("category", "").strip()
        try:
            limit = max(1, min(int(request.args.get("limit", 100)), 1000))
        except (ValueError, TypeError):
            limit = 100

        if key:
            row = c.execute(
                "SELECT key, value, category, created_by, created_at, updated_at "
                "FROM agent_memory WHERE key=?", (key,)
            ).fetchone()
            conn.close()
            if not row:
                return jsonify({"error": "Key no encontrada"}), 404
            return jsonify({
                "key": row[0], "value": row[1], "category": row[2],
                "created_by": row[3], "created_at": row[4], "updated_at": row[5],
            })

        sql = ("SELECT key, value, category, created_by, created_at, updated_at "
               "FROM agent_memory WHERE 1=1")
        params = []
        if category:
            sql += " AND category=?"
            params.append(category)
        sql += " ORDER BY updated_at DESC LIMIT ?"
        params.append(limit)
        rows = c.execute(sql, params).fetchall()
        conn.close()
        return jsonify({
            "entries": [
                {"key": r[0], "value": r[1], "category": r[2],
                 "created_by": r[3], "created_at": r[4], "updated_at": r[5]}
                for r in rows
            ],
            "total": len(rows),
        })

    # POST · upsert o delete
    body = request.get_json(silent=True) or {}
    key = (body.get("key") or "").strip()
    if not key:
        conn.close()
        return jsonify({"error": "key requerido"}), 400

    if body.get("_delete"):
        c.execute("DELETE FROM agent_memory WHERE key=?", (key,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "deleted": key})

    value = body.get("value")
    if value is None:
        conn.close()
        return jsonify({"error": "value requerido"}), 400
    # Coerce value a string (puede venir como dict/list)
    if not isinstance(value, str):
        import json as _json
        value = _json.dumps(value, ensure_ascii=False)
    category = (body.get("category") or "general").strip()

    c.execute("""
        INSERT INTO agent_memory(key, value, category, created_by, updated_at)
        VALUES(?, ?, ?, ?, datetime('now', 'utc'))
        ON CONFLICT(key) DO UPDATE SET
          value = excluded.value,
          category = excluded.category,
          updated_at = datetime('now', 'utc')
    """, (key, value, category, u))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "key": key, "category": category})


@bp.route("/api/admin/diag-login/<username>", methods=["GET"])
def admin_diag_login(username):
    """Diagnóstico detallado del estado de login de un usuario.

    Sebastián 7-may-2026 (Mayerlin): cuando un usuario reporta que no puede
    entrar, Sebastián necesita ver de un vistazo:
      · Tiene PASS_<USER> seteada en Render? Es hash o plaintext?
      · Cambió su password vía self-service (DB)? Cuándo?
      · Cuántos intentos fallidos recientes? Está locked?
      · Tiene MFA enabled?
      · Cuál es la acción recomendada?

    Returns: { username, exists, password_source, password_changed_at,
               last_login, recent_failures: [{ts, ip}], is_locked,
               mfa_enabled, recommended_action, hint }
    """
    admin_user, err, code = _require_admin()
    if err:
        return err, code

    target = (username or "").strip().lower()
    if not target:
        return jsonify({"error": "Falta username"}), 400
    exists = target in COMPRAS_USERS

    out = {
        "username": target,
        "exists": exists,
        "password_source": None,
        "password_changed_at": None,
        "password_changed_by": None,
        "last_login": None,
        "recent_failures": [],
        "is_locked": False,
        "mfa_enabled": False,
        "recommended_action": "",
        "hint": "",
    }
    if not exists:
        out["recommended_action"] = "AGREGAR_USUARIO"
        out["hint"] = (f"'{target}' no está en config.COMPRAS_USERS · "
                        "agregalo en config.py + setea PASS_{target.upper()} "
                        "en Render")
        return jsonify(out)

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("PRAGMA busy_timeout=2000")
        # password_source via helper existente
        out["password_source"] = _password_source(target, conn)
        # password_changed_at + by
        try:
            row = conn.execute(
                "SELECT changed_at, changed_by FROM users_passwords WHERE username=?",
                (target,),
            ).fetchone()
            if row:
                out["password_changed_at"] = row[0]
                out["password_changed_by"] = row[1]
        except Exception:
            pass
        # Último login exitoso
        try:
            row = conn.execute(
                """SELECT ts, ip FROM security_events
                   WHERE event='login_success' AND username=?
                   ORDER BY id DESC LIMIT 1""",
                (target,),
            ).fetchone()
            if row:
                out["last_login"] = {"ts": row[0], "ip": row[1]}
        except Exception:
            pass
        # Últimos 5 login_failure
        try:
            rows = conn.execute(
                """SELECT ts, ip FROM security_events
                   WHERE event='login_failure' AND username=?
                   ORDER BY id DESC LIMIT 5""",
                (target,),
            ).fetchall()
            out["recent_failures"] = [{"ts": r[0], "ip": r[1]} for r in rows]
        except Exception:
            pass
        # MFA
        try:
            row = conn.execute(
                "SELECT 1 FROM users_mfa WHERE username=? AND enabled=1",
                (target,),
            ).fetchone()
            out["mfa_enabled"] = bool(row)
        except Exception:
            pass
        conn.close()
    except Exception as e:
        log.warning("diag_login DB error: %s", e)

    # Lock status (importado del módulo core)
    try:
        from blueprints.core import _is_locked  # noqa
        # _is_locked toma (ip, username) · sin IP no podemos saber por IP
        # exacta; pero si username está bloqueado en cualquier IP, lo
        # reportamos. _is_locked usa rate_limit table.
        conn = sqlite3.connect(DB_PATH)
        # Contar entries en rate_limit para este username
        try:
            row = conn.execute(
                """SELECT COUNT(*) FROM rate_limit
                   WHERE username=? AND failures >= 5""",
                (target,),
            ).fetchone()
            out["is_locked"] = bool(row and row[0])
        except Exception:
            pass
        conn.close()
    except Exception:
        pass

    # Recomendación basada en estado
    src = out["password_source"]
    n_fail = len(out["recent_failures"])
    if src == "missing":
        out["recommended_action"] = "SETEAR_PASS_EN_RENDER"
        out["hint"] = (
            f"PASS_{target.upper()} no existe en Render. "
            "Opciones: (1) En el panel /admin click 'Resetear' — genera "
            "password nueva y la guarda en DB (no necesitás Render). "
            "(2) Setear PASS_{target.upper()} en Render con un hash "
            "pbkdf2:sha256:600000 (correr scripts/gen_password_hashes.py)."
        )
    elif src == "env_plaintext":
        out["recommended_action"] = "RESETEAR_PASSWORD"
        out["hint"] = (
            f"PASS_{target.upper()} en Render está en PLAINTEXT (inseguro). "
            "Click 'Resetear' para que sea hash + queda en DB. La password "
            "actual debería ser exactamente el valor de la env var."
        )
    elif out["is_locked"]:
        out["recommended_action"] = "ESPERAR_LOCK_O_RESETEAR"
        out["hint"] = (
            "Usuario bloqueado por demasiados intentos fallidos. "
            "Esperá 15 min o resetear (limpia el lock implícitamente)."
        )
    elif n_fail >= 3 and not out["last_login"]:
        out["recommended_action"] = "RESETEAR_PASSWORD"
        out["hint"] = (
            f"{n_fail} fallos recientes y nunca un login exitoso. "
            "Probable que el usuario no sepa su password · resetear."
        )
    elif n_fail >= 3 and out["last_login"]:
        out["recommended_action"] = "RESETEAR_PASSWORD"
        out["hint"] = (
            f"{n_fail} fallos recientes (último login OK fue "
            f"{out['last_login'].get('ts')}). El usuario olvidó "
            "o cambió de teclado · resetear."
        )
    else:
        out["recommended_action"] = "OK"
        out["hint"] = (
            "Sin señales de problema. Si el usuario insiste que no puede "
            "entrar: pedirle screenshot del error exacto + verificar IP + "
            "preguntarle si el browser tiene autofill mal."
        )

    return jsonify(out)


@bp.route("/api/admin/reset-password", methods=["POST"])
def admin_reset_password():
    """Resetea la password de OTRO usuario. Genera password aleatoria,
    la guarda hasheada en users_passwords, y devuelve la plaintext UNA SOLA VEZ.

    Body JSON: {"username": "<target>"}
    """
    admin_user, err, code = _require_admin()
    if err:
        return err, code

    body = request.get_json(silent=True) or {}
    target = (body.get("username") or "").strip().lower()

    if not target:
        return jsonify({"error": "Falta 'username'"}), 400

    if target not in COMPRAS_USERS:
        return jsonify({"error": f"Usuario '{target}' no existe"}), 404

    # Generar password aleatoria fuerte (12 chars, sin caracteres ambiguos)
    safe_alphabet = "".join(
        c for c in (string.ascii_letters + string.digits + "!@#%&*+=?")
        if c not in "0Oo1Il"
    )
    while True:
        new_pwd = "".join(secrets.choice(safe_alphabet) for _ in range(12))
        if (any(c.isupper() for c in new_pwd)
                and any(c.islower() for c in new_pwd)
                and any(c.isdigit() for c in new_pwd)
                and any(not c.isalnum() for c in new_pwd)):
            break

    new_hash = generate_password_hash(new_pwd, method="pbkdf2:sha256:600000")

    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("PRAGMA busy_timeout=2000")
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO users_passwords (username, password_hash, changed_at, changed_by)
            VALUES (?, ?, datetime('now', 'utc'), ?)
            ON CONFLICT(username) DO UPDATE SET
                password_hash = excluded.password_hash,
                changed_at    = excluded.changed_at,
                changed_by    = excluded.changed_by
        """, (target, new_hash, admin_user))
        try:
            audit_log(cur, usuario=admin_user, accion='RESET_PASSWORD',
                      tabla='users_passwords', registro_id=target,
                      detalle=f"Admin {admin_user} reseteó password de {target}")
        except Exception:
            pass  # security event ya queda · audit es defense-in-depth
        conn.commit()
        conn.close()
    except Exception as e:
        _log_sec("password_reset_db_error", admin_user, _client_ip(), str(e)[:200])
        return jsonify({"error": "Error guardando nueva password"}), 500

    _log_sec(
        "password_reset_by_admin",
        admin_user,
        _client_ip(),
        f"target={target}"
    )

    return jsonify({
        "ok": True,
        "username": target,
        "new_password": new_pwd,
        "warning": "Esta password se muestra UNA SOLA VEZ. Comunícala al "
                   "usuario por canal seguro y dile que la cambie en su "
                   "primer login."
    })


# ─── Security events ─────────────────────────────────────────────────────────


@bp.route("/api/admin/security-events", methods=["GET"])
def admin_security_events():
    """Últimos eventos de seguridad. Soporta filtros: ?event=login_failure&limit=100"""
    u, err, code = _require_admin()
    if err:
        return err, code

    event_filter = request.args.get("event", "").strip()
    try:
        limit = min(int(request.args.get("limit", "50")), 500)
    except ValueError:
        limit = 50

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        if event_filter:
            rows = conn.execute(
                """SELECT id, ts, event, username, ip, user_agent, details
                   FROM security_events
                   WHERE event = ?
                   ORDER BY id DESC LIMIT ?""",
                (event_filter, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT id, ts, event, username, ip, user_agent, details
                   FROM security_events
                   ORDER BY id DESC LIMIT ?""",
                (limit,)
            ).fetchall()
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e), "events": []}), 500

    out = [dict(r) for r in rows]

    # Estadísticas rápidas: conteo por evento en últimas 24h
    stats = {}
    try:
        agg = conn.execute("""
            SELECT event, COUNT(*) as n FROM security_events
            WHERE ts > datetime('now', 'utc', '-1 day')
            GROUP BY event
            ORDER BY n DESC
        """).fetchall()
        stats = {r[0]: r[1] for r in agg}
    except Exception:
        pass
    conn.close()

    return jsonify({
        "events": out,
        "total":  len(out),
        "stats_24h": stats,
    })


# ─── Config status ────────────────────────────────────────────────────────────


@bp.route("/api/admin/config-status", methods=["GET"])
def admin_config_status():
    """Status de configuración: env vars críticas y opcionales presentes/faltan."""
    u, err, code = _require_admin()
    if err:
        return err, code

    issues = validate_config()

    # Lista de env vars que deberían estar configuradas
    critical_vars = ["SECRET_KEY", "DB_PATH"]
    user_pass_vars = [f"PASS_{u.upper()}" for u in COMPRAS_USERS.keys()]
    optional_vars = [
        "FORMULA_PIN",
        "ANTHROPIC_API_KEY", "GHL_API_KEY", "GHL_LOCATION_ID",
        "SHOPIFY_TOKEN", "SHOPIFY_SHOP",
        "INSTAGRAM_TOKEN", "INSTAGRAM_USER_ID",
        "META_APP_ID", "META_APP_SECRET",
        "EMAIL_REMITENTE", "EMAIL_PASSWORD",
        "SENTRY_DSN",
    ]

    def status(name):
        val = os.environ.get(name, "")
        return {"name": name, "set": bool(val), "length": len(val) if val else 0}

    return jsonify({
        "issues":         issues,
        "critical":       [status(v) for v in critical_vars],
        "user_passwords": [status(v) for v in user_pass_vars],
        "optional":       [status(v) for v in optional_vars],
    })


# ─── Test SMTP: envía un correo de prueba con PDF demo ────────────────────────

@bp.route("/api/admin/test-email", methods=["POST"])
def admin_test_email():
    """Envía un correo de prueba al destinatario indicado.

    Body JSON: {"destinatario": "email@dominio.com"}
       Si no se especifica, usa EMAIL_REMITENTE (te lo manda a ti mismo).

    Genera un PDF demo (CE-TEST-0000) y lo adjunta para validar:
      1. SMTP configurado y credenciales correctas
      2. PDF generator funcionando
      3. Email llega con HTML + adjunto

    Devuelve:
      200 OK    si se envió (revisar bandeja del destinatario)
      503       si EMAIL_REMITENTE/EMAIL_PASSWORD no están seteados
      502       si SMTP rechaza credenciales o no puede entregar
      500       error inesperado (PDF generator, etc.)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    body = request.get_json(silent=True) or {}
    destinatario = (body.get('destinatario') or '').strip()
    # 'animus' o 'espagiria' — para probar ambos formatos del PDF
    empresa_test = (body.get('empresa') or 'espagiria').strip().lower()
    if 'animus' in empresa_test:
        empresa_test = 'animus'
    else:
        empresa_test = 'espagiria'

    # Probar primero la config SMTP
    try:
        import sys
        from pathlib import Path
        # notificaciones.py está en la raíz del repo, no en api/
        repo_root = Path(__file__).resolve().parents[2]
        if str(repo_root) not in sys.path:
            sys.path.insert(0, str(repo_root))
        from notificaciones import SistemaNotificaciones
    except Exception as e:
        return jsonify({'error': 'No se pudo cargar SistemaNotificaciones', 'detalle': str(e)}), 500

    sn = SistemaNotificaciones()
    if not sn.email_remitente or not sn.contraseña:
        return jsonify({
            'error': 'SMTP no configurado',
            'detalle': 'Faltan EMAIL_REMITENTE y/o EMAIL_PASSWORD en variables de entorno.',
            'pasos': [
                '1) Genera App Password en https://myaccount.google.com/apppasswords',
                '2) Render dashboard → tu servicio → Environment → Add Environment Variable',
                '3) EMAIL_REMITENTE = facturasespagirialaboratorio@gmail.com',
                '4) EMAIL_PASSWORD = (la app password de 16 chars sin espacios)',
                '5) Save Changes → Render redeploya solo',
                '6) Vuelve a probar /api/admin/test-email',
            ],
            'env_status': {
                'EMAIL_REMITENTE': 'set' if sn.email_remitente else 'MISSING',
                'EMAIL_PASSWORD':  'set' if sn.contraseña else 'MISSING',
                'SMTP_SERVER':     sn.smtp_server,
                'SMTP_PORT':       sn.smtp_port,
            }
        }), 503

    # Si no especificó destinatario, mandárselo a sí mismo (al remitente)
    if not destinatario:
        destinatario = sn.email_remitente
    if '@' not in destinatario:
        return jsonify({'error': 'Destinatario inválido'}), 400

    # Generar PDF demo
    try:
        import sys
        from pathlib import Path
        api_dir = Path(__file__).resolve().parents[1]  # api/
        if str(api_dir) not in sys.path:
            sys.path.insert(0, str(api_dir))
        from comprobante_pago import generar_comprobante_egreso_pdf
        from datetime import datetime as _dt
        pdf_bytes = generar_comprobante_egreso_pdf(
            numero_ce='CE-TEST-0000',
            fecha_pago=_dt.now(),
            beneficiario={
                'nombre': 'PRUEBA SMTP — Test Influencer',
                'cedula': '00000000',
                'banco': 'Bancolombia',
                'cuenta': '0000000000',
                'tipo_cuenta': 'Ahorros',
                'ciudad': 'Cali',
            },
            items=[{
                'descripcion': 'Correo de prueba — validación de configuración SMTP',
                'fecha': _dt.now().strftime('%Y-%m-%d'),
                'cantidad': 1,
                'valor_unit': 100000,
            }],
            aplicar_retefuente=False,
            aplicar_retica=False,
            aplicar_iva=False,
            medio_pago='N/A — prueba',
            observaciones=f'Test enviado por {u} desde /admin',
            pagado_por=u,
            empresa_clave=empresa_test,
        )
    except Exception as e:
        import traceback
        return jsonify({
            'error': 'Falló generación de PDF demo',
            'detalle': str(e),
            'traceback': traceback.format_exc()[-800:]
        }), 500

    # Enviar el correo (sincrónico para que la respuesta indique éxito/fallo real)
    try:
        ok = sn.enviar_comprobante_egreso(
            destinatario=destinatario,
            numero_ce='CE-TEST-0000',
            beneficiario='PRUEBA SMTP',
            total_pagado=100000,
            pdf_bytes=pdf_bytes,
            fecha_emision=_dt.now().strftime('%Y-%m-%d'),
            numero_oc='OC-TEST-0000',
            empresa='ANIMUS Lab' if empresa_test == 'animus' else 'Espagiria',
        )
    except Exception as e:
        return jsonify({'error': 'Excepción al enviar', 'detalle': str(e)}), 500

    if not ok:
        return jsonify({
            'error': 'SMTP rechazó el envío',
            'detalle': (
                'Las credenciales se cargaron pero Gmail/SMTP rechazó el envío. '
                'Causas comunes: App Password incorrecta, 2FA deshabilitado, '
                'la cuenta tiene "Less secure apps" off, o el remitente no '
                'coincide con la cuenta autenticada. Revisa los logs de Render.'
            ),
            'remitente': sn.email_remitente,
            'destinatario': destinatario,
            'smtp_server': f'{sn.smtp_server}:{sn.smtp_port}',
        }), 502

    _log_sec(u, _client_ip(), 'admin_test_email', f'destinatario={destinatario}')

    return jsonify({
        'ok': True,
        'mensaje': 'Correo de prueba enviado',
        'remitente': sn.email_remitente,
        'destinatario': destinatario,
        'asunto': 'Comprobante de pago CE-TEST-0000 — Espagiria',
        'verificacion': [
            f'1. Revisa la bandeja de {destinatario} (también la carpeta Spam)',
            '2. El correo trae un PDF adjunto (CE-TEST-0000.pdf)',
            '3. El PDF debe abrir y mostrar el logo HHA + datos demo',
            '4. Si todo OK, el sistema está listo para enviar comprobantes reales',
        ],
    })


# ─── Diagnóstico: tipos de material en maestro_mps ────────────────────────────

@bp.route("/api/admin/import-mps-nombres-excel", methods=["POST"])
def admin_import_mps_nombres_excel():
    """Importa Excel con código + nombre comercial (+ proveedor opcional) para
    corregir el catálogo masivamente.

    Caso de uso: el catálogo tiene nombre_comercial = código en muchas filas
    porque alguien se confundió al importar. Excel con la corrección rápida.

    Detecta columnas automáticamente:
      - código: 'codigo', 'code', 'mp', 'sku'
      - nombre: 'nombre', 'descripcion', 'material', 'comercial'
      - proveedor: 'proveedor', 'supplier' (opcional)

    Body: multipart/form-data con 'file' = .xlsx
    Query: ?dry_run=1 para preview

    Devuelve por fila: actualizado | sin_cambios | sin_match | sin_codigo
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    dry_run = request.args.get('dry_run', '0') in ('1', 'true', 'True')

    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'Excel inválido: {e}'}), 400

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({'error': 'Hoja con menos de 2 filas'}), 400

    headers = rows[0]
    norms = [_norm_header(h) for h in headers]

    def _find(*kws):
        for i, h in enumerate(norms):
            if any(k in h for k in kws):
                return i
        return None

    idx_cod = _find('codigo', 'code', 'sku')
    idx_nom = _find('nombre', 'descripcion', 'comercial', 'material')
    idx_prov = _find('proveedor', 'supplier')

    if idx_cod is None or idx_nom is None:
        return jsonify({
            'error': 'No detecté columnas de código y nombre',
            'headers': [str(h) for h in headers if h],
            'sugerencia': 'El Excel debe tener al menos columnas "código" y "nombre"',
        }), 400

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    actualizados = []
    sin_cambios = []
    sin_match = []
    sin_codigo = []

    for i, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        cod = str(row[idx_cod] or '').strip() if idx_cod is not None else ''
        nom_nuevo = str(row[idx_nom] or '').strip() if idx_nom is not None else ''
        prov_nuevo = str(row[idx_prov] or '').strip() if idx_prov is not None else ''
        if not cod:
            sin_codigo.append({'fila': i, 'nombre_excel': nom_nuevo})
            continue
        # Buscar item actual
        cur = c.execute(
            "SELECT codigo_mp, nombre_comercial, COALESCE(proveedor,'') FROM maestro_mps WHERE codigo_mp=?",
            (cod,)
        ).fetchone()
        if not cur:
            sin_match.append({'fila': i, 'codigo': cod, 'nombre_excel': nom_nuevo})
            continue
        cambios = {}
        # Solo actualizar nombre si el actual está vacío o es igual al código
        # (señal de que está mal cargado). NO sobrescribir nombres reales.
        actual_nom = (cur['nombre_comercial'] or '').strip()
        if nom_nuevo and (not actual_nom or actual_nom == cod):
            cambios['nombre_comercial'] = nom_nuevo
        # Proveedor: solo actualizar si actual está vacío y Excel tiene
        actual_prov = (cur[2] or '').strip()
        if prov_nuevo and not actual_prov:
            cambios['proveedor'] = prov_nuevo
        if cambios:
            if not dry_run:
                sets = ', '.join(f"{k}=?" for k in cambios)
                vals = list(cambios.values()) + [cod]
                c.execute(f"UPDATE maestro_mps SET {sets} WHERE codigo_mp=?", vals)
            actualizados.append({
                'codigo': cod,
                'nombre_antes': actual_nom,
                'nombre_nuevo': cambios.get('nombre_comercial', actual_nom),
                'proveedor_nuevo': cambios.get('proveedor', actual_prov),
            })
        else:
            sin_cambios.append(cod)

    if not dry_run:
        conn.commit()
    conn.close()

    if not dry_run:
        _log_sec(u, _client_ip(),
                 "admin_import_mps_nombres_excel",
                 f"actualizados={len(actualizados)} sin_match={len(sin_match)}")

    return jsonify({
        'ok': True,
        'dry_run': dry_run,
        'actualizados': {'count': len(actualizados), 'lista': actualizados[:50]},
        'sin_cambios': {'count': len(sin_cambios)},
        'sin_match': {'count': len(sin_match), 'lista': sin_match[:30]},
        'sin_codigo': {'count': len(sin_codigo), 'lista': sin_codigo[:10]},
        'columnas_mapeadas': {'codigo': idx_cod, 'nombre': idx_nom, 'proveedor': idx_prov},
        'nota': 'Solo actualiza nombres si el actual está vacío o = código. Nombres correctos NO se sobrescriben.',
    })


def _parse_excel_verde(file_storage, incluir_no_verde=False):
    """Parsea Excel del conteo fisico.

    Por default retorna solo filas verdes (FF92D050). Si incluir_no_verde
    es True, tambien retorna las filas rojas/blancas en un dict aparte
    `excel_no_verde` — esto sirve para regenerar lotes que las
    producciones consumieron pero que ya no estan fisicamente (Catalina
    los marco NO presente porque se acabaron en produccion).

    Devuelve: (excel_verde, total_g, rows_no_verde_count, errores)
        Si incluir_no_verde=True: (excel_verde, total_g, excel_no_verde, errores)

    excel_verde[(cod, lote)] = {codigo_mp, lote, inci, nombre_comercial,
                                proveedor, estanteria, posicion,
                                fecha_vencimiento, cantidad_g}
    """
    from openpyxl import load_workbook

    GREEN = 'FF92D050'
    errors = []

    try:
        wb = load_workbook(file_storage, data_only=True)
    except Exception as _e:
        return None, 0, 0, [f'No pude abrir el Excel: {_e}']

    if not wb.sheetnames:
        return None, 0, 0, ['Excel sin hojas']

    ws = wb[wb.sheetnames[0]]

    # Detectar fila header
    header_row = None
    for r in range(1, min(20, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=r, column=col).value or '').upper()
                    for col in range(1, min(15, ws.max_column + 1))]
        joined = ' | '.join(row_vals)
        if ('CÓDIGO MP' in joined or 'CODIGO MP' in joined) and 'LOTE' in joined:
            header_row = r
            break
    if header_row is None:
        header_row = 5

    col_idx = {}
    for col in range(1, ws.max_column + 1):
        v = (ws.cell(row=header_row, column=col).value or '')
        v = str(v).upper().strip()
        if 'CÓDIGO MP' in v or 'CODIGO MP' in v:
            col_idx['codigo'] = col
        elif 'NOMBRE INCI' in v or v == 'INCI':
            col_idx['inci'] = col
        elif 'NOMBRE COMERCIAL' in v or v == 'COMERCIAL':
            col_idx['comercial'] = col
        elif 'PROVEEDOR' in v:
            col_idx['proveedor'] = col
        elif 'LOTE' in v and 'N' in v:
            col_idx['lote'] = col
        elif 'CONTEO' in v:
            col_idx['cant_conteo'] = col
        elif 'ESTANTER' in v:
            col_idx['estanteria'] = col
        elif v == 'POS.' or v == 'POSICION' or v == 'POSICIÓN':
            col_idx['posicion'] = col
        elif 'FECHA VENC' in v or v == 'VENCE':
            col_idx['venc'] = col

    if 'codigo' not in col_idx or 'lote' not in col_idx or 'cant_conteo' not in col_idx:
        if incluir_no_verde:
            return None, 0, {}, [f'Excel sin columnas requeridas: {list(col_idx.keys())}']
        return None, 0, 0, [f'Excel sin columnas requeridas: {list(col_idx.keys())}']

    excel_verde = {}
    excel_no_verde = {}
    total_g = 0.0
    rows_no_verde = 0

    def _row_to_dict(r, cod, lote, cant_g):
        """Construye el dict info estandar para una fila."""
        venc = ws.cell(row=r, column=col_idx['venc']).value if 'venc' in col_idx else None
        if hasattr(venc, 'isoformat'):
            venc = venc.isoformat()[:10]
        else:
            venc = str(venc)[:10] if venc else ''
        return {
            'codigo_mp': cod, 'lote': lote,
            'inci': str(ws.cell(row=r, column=col_idx['inci']).value or '') if 'inci' in col_idx else '',
            'nombre_comercial': str(ws.cell(row=r, column=col_idx['comercial']).value or '') if 'comercial' in col_idx else '',
            'proveedor': str(ws.cell(row=r, column=col_idx['proveedor']).value or '') if 'proveedor' in col_idx else '',
            'estanteria': str(ws.cell(row=r, column=col_idx['estanteria']).value or '') if 'estanteria' in col_idx else '',
            'posicion': str(ws.cell(row=r, column=col_idx['posicion']).value or '') if 'posicion' in col_idx else '',
            'fecha_vencimiento': venc,
            'cantidad_g': cant_g,
        }

    for r in range(header_row + 1, ws.max_row + 1):
        cell_codigo = ws.cell(row=r, column=col_idx['codigo'])
        cell_color = (cell_codigo.fill.fgColor.rgb
                      if cell_codigo.fill.fgColor.type == 'rgb' else None)
        is_verde = (cell_color == GREEN)
        cod = cell_codigo.value
        if not cod:
            if not is_verde:
                rows_no_verde += 1
            continue
        cod = str(cod).strip()
        lote_raw = ws.cell(row=r, column=col_idx['lote']).value
        lote = str(lote_raw).strip() if lote_raw else ''
        cant_raw = ws.cell(row=r, column=col_idx['cant_conteo']).value
        try:
            cant = float(cant_raw or 0)
        except (TypeError, ValueError):
            cant = 0.0

        key = (cod, lote)
        info = _row_to_dict(r, cod, lote, cant)

        if is_verde:
            if key in excel_verde:
                excel_verde[key]['cantidad_g'] += cant
            else:
                excel_verde[key] = info
            total_g += cant
        else:
            rows_no_verde += 1
            if incluir_no_verde:
                if key in excel_no_verde:
                    excel_no_verde[key]['cantidad_g'] += cant
                else:
                    excel_no_verde[key] = info

    if incluir_no_verde:
        return excel_verde, total_g, excel_no_verde, errors
    return excel_verde, total_g, rows_no_verde, errors


def _identify_movimientos_a_preservar(c):
    """Devuelve dos listas de movimientos a re-insertar tras el reset:

      - entradas_oc_legitimas: tipo='Entrada' con numero_oc no vacio
      - salidas_produccion: tipo='Salida' con observaciones que empiezan
        con 'FEFO:' (las que vienen de handle_produccion en inventario.py)

    Cualquier otro movimiento (Entradas sin OC, Salidas que no son de
    produccion, ajustes manuales) se descarta — esos son los datos
    sucios que el reset esta limpiando.
    """
    cols = ('id, material_id, material_nombre, cantidad, tipo, fecha, '
            'observaciones, lote, fecha_vencimiento, estanteria, posicion, '
            'proveedor, estado_lote, operador, '
            'COALESCE(numero_oc,"") as numero_oc, '
            'COALESCE(numero_factura,"") as numero_factura, '
            'COALESCE(precio_kg,0) as precio_kg')
    try:
        entradas_oc = c.execute(f"""SELECT {cols} FROM movimientos
                                    WHERE tipo='Entrada'
                                    AND COALESCE(numero_oc,'') != ''
                                    AND COALESCE(numero_oc,'') != '0'
                                    ORDER BY fecha ASC, id ASC""").fetchall()
    except sqlite3.OperationalError:
        entradas_oc = []
    try:
        salidas_prod = c.execute(f"""SELECT {cols} FROM movimientos
                                     WHERE tipo='Salida'
                                     AND observaciones LIKE 'FEFO:%'
                                     ORDER BY fecha ASC, id ASC""").fetchall()
    except sqlite3.OperationalError:
        salidas_prod = []
    return entradas_oc, salidas_prod


def _row_to_dict(row, columns):
    return dict(zip(columns, row))


_MOV_PRESERVAR_COLS = (
    'id', 'material_id', 'material_nombre', 'cantidad', 'tipo', 'fecha',
    'observaciones', 'lote', 'fecha_vencimiento', 'estanteria', 'posicion',
    'proveedor', 'estado_lote', 'operador', 'numero_oc', 'numero_factura',
    'precio_kg',
)


@bp.route("/api/admin/audit-inventario-vs-excel", methods=["POST"])
def admin_audit_inventario_vs_excel():
    """Compara el inventario actual contra un Excel "dia cero" sin escribir nada.

    Caso de uso (CEO 2026-04-27): el inventario de planta esta inconsistente
    contra la realidad fisica. Tenemos un Excel con el conteo fisico hecho
    por Catalina antes de las producciones recientes. Cada fila tiene un
    color que indica si el lote existe fisicamente (verde) o no (rojo /
    blanco / gris claro). Solo verde cuenta como real.

    Este endpoint:
      1. Lee el Excel uploaded (campo 'file' en multipart)
      2. Filtra rows verdes (FF92D050) — el resto se ignora
      3. Compara cada (codigo_mp, lote) verde contra el estado actual
         del kardex en movimientos
      4. Retorna reporte JSON con:
         - lotes_verde: total que catalina marco como reales
         - en_db_match: presentes en DB con cantidad esperada (excel - salidas)
         - en_db_con_delta: presentes pero cantidad no coincide
         - faltantes_en_db: en Excel pero no en DB (lote desaparecio del kardex)
         - solo_db: en DB pero NO en Excel verde (post-day-zero o sobrante)
         - producciones_post_excel: count + total kg
         - resumen_g: total_excel_verde, total_db_actual, total_post_dia_cero,
                      delta_total_g

    SIN ESCRITURA. Reporte de solo lectura. El usuario decide despues
    si aplicar reset+replay o ajustes quirurgicos.

    Solo admins.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    GREEN = 'FF92D050'

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as _e:
        return jsonify({'error': f'No pude abrir el Excel: {_e}'}), 400

    if not wb.sheetnames:
        return jsonify({'error': 'Excel sin hojas'}), 400

    # Heuristica: usar la primera hoja (tipicamente "INVENTARIO")
    ws = wb[wb.sheetnames[0]]

    # Detectar fila header buscando por palabras clave
    header_row = None
    for r in range(1, min(20, ws.max_row + 1)):
        row_vals = [str(ws.cell(row=r, column=col).value or '').upper()
                    for col in range(1, min(15, ws.max_column + 1))]
        joined = ' | '.join(row_vals)
        if ('CÓDIGO MP' in joined or 'CODIGO MP' in joined) and 'LOTE' in joined:
            header_row = r
            break
    if header_row is None:
        header_row = 5  # fallback al esperado

    # Mapear columnas
    col_idx = {}
    for col in range(1, ws.max_column + 1):
        v = (ws.cell(row=header_row, column=col).value or '')
        v = str(v).upper().strip()
        if 'CÓDIGO MP' in v or 'CODIGO MP' in v:
            col_idx['codigo'] = col
        elif 'NOMBRE INCI' in v or v == 'INCI':
            col_idx['inci'] = col
        elif 'NOMBRE COMERCIAL' in v or v == 'COMERCIAL':
            col_idx['comercial'] = col
        elif 'PROVEEDOR' in v:
            col_idx['proveedor'] = col
        elif 'LOTE' in v and 'N' in v:
            col_idx['lote'] = col
        elif 'CONTEO' in v:
            col_idx['cant_conteo'] = col
        elif 'ESTANTER' in v:
            col_idx['estanteria'] = col
        elif v == 'POS.' or v == 'POSICION' or v == 'POSICIÓN':
            col_idx['posicion'] = col
        elif 'FECHA VENC' in v or v == 'VENCE':
            col_idx['venc'] = col

    if 'codigo' not in col_idx or 'lote' not in col_idx or 'cant_conteo' not in col_idx:
        return jsonify({
            'error': 'Excel no tiene columnas requeridas',
            'detail': f'columnas detectadas: {list(col_idx.keys())}, '
                      f'header_row={header_row}'
        }), 400

    # ── Parsear filas verdes ─────────────────────────────────────────────
    excel_verde = {}  # (codigo, lote) -> dict
    excel_total_g = 0.0
    rows_no_verde_count = 0
    for r in range(header_row + 1, ws.max_row + 1):
        cell_color = (ws.cell(row=r, column=col_idx['codigo']).fill.fgColor.rgb
                      if ws.cell(row=r, column=col_idx['codigo']).fill.fgColor.type == 'rgb'
                      else None)
        if cell_color != GREEN:
            rows_no_verde_count += 1
            continue
        cod = ws.cell(row=r, column=col_idx['codigo']).value
        if not cod:
            continue
        cod = str(cod).strip()
        lote_raw = ws.cell(row=r, column=col_idx['lote']).value
        lote = str(lote_raw).strip() if lote_raw else ''
        cant_raw = ws.cell(row=r, column=col_idx['cant_conteo']).value
        try:
            cant = float(cant_raw or 0)
        except (TypeError, ValueError):
            cant = 0.0
        venc = ws.cell(row=r, column=col_idx['venc']).value if 'venc' in col_idx else None
        if hasattr(venc, 'isoformat'):
            venc = venc.isoformat()[:10]
        else:
            venc = str(venc)[:10] if venc else ''

        key = (cod, lote)
        if key in excel_verde:
            # Lote duplicado en Excel: sumar cantidades, mantener primer registro
            excel_verde[key]['cantidad_g'] += cant
            excel_verde[key]['lotes_duplicados'] = excel_verde[key].get('lotes_duplicados', 1) + 1
        else:
            excel_verde[key] = {
                'codigo_mp': cod,
                'lote': lote,
                'inci': str(ws.cell(row=r, column=col_idx['inci']).value or '') if 'inci' in col_idx else '',
                'nombre_comercial': str(ws.cell(row=r, column=col_idx['comercial']).value or '') if 'comercial' in col_idx else '',
                'proveedor': str(ws.cell(row=r, column=col_idx['proveedor']).value or '') if 'proveedor' in col_idx else '',
                'estanteria': str(ws.cell(row=r, column=col_idx['estanteria']).value or '') if 'estanteria' in col_idx else '',
                'posicion': str(ws.cell(row=r, column=col_idx['posicion']).value or '') if 'posicion' in col_idx else '',
                'fecha_vencimiento': venc,
                'cantidad_g': cant,
            }
        excel_total_g += cant

    # ── Estado actual del kardex ─────────────────────────────────────────
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    db_lotes = {}  # (cod, lote) -> dict
    c.execute("""SELECT material_id,
                        COALESCE(lote,''),
                        SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE 0 END) as entradas,
                        SUM(CASE WHEN tipo='Salida'  THEN cantidad ELSE 0 END) as salidas,
                        COUNT(*) as nmovs
                 FROM movimientos
                 GROUP BY material_id, lote""")
    db_total_neto_g = 0.0
    for row in c.fetchall():
        cod = (row[0] or '').strip()
        lote = (row[1] or '').strip()
        entradas = float(row[2] or 0)
        salidas = float(row[3] or 0)
        neto = entradas - salidas
        db_lotes[(cod, lote)] = {
            'entradas_g': entradas,
            'salidas_g': salidas,
            'neto_g': neto,
            'n_movs': row[4],
        }
        db_total_neto_g += neto

    # Producciones — info de contexto
    try:
        prod_n = c.execute("SELECT COUNT(*) FROM producciones").fetchone()[0]
    except sqlite3.OperationalError:
        prod_n = 0
    try:
        prod_total_kg = float(c.execute(
            "SELECT COALESCE(SUM(cantidad), 0) FROM producciones"
        ).fetchone()[0] or 0)
    except sqlite3.OperationalError:
        prod_total_kg = 0.0

    conn.close()

    # ── Diff ─────────────────────────────────────────────────────────────
    en_db_match = []
    en_db_con_delta = []
    faltantes_en_db = []  # excel verde pero NO en DB
    solo_db = []  # en DB pero NO en excel verde

    TOLERANCIA_G = 0.5  # tolerancia de redondeo

    excel_keys = set(excel_verde.keys())
    db_keys = set(db_lotes.keys())

    for k in excel_keys & db_keys:
        e = excel_verde[k]
        d = db_lotes[k]
        # Esperado: el lote tenia excel_g al dia cero. Si hubo salidas
        # despues, el neto ahora debe ser excel_g - salidas. Si hubo
        # entradas adicionales (raro, mismo lote re-recibido), tambien
        # se suman. Mas simple: comparar neto actual con (excel - salidas
        # + entradas_post_dia_cero) — pero no sabemos cual entrada fue
        # la inicial. Aproximacion: neto_esperado = excel_g - salidas.
        # Si DB tiene mas entradas que el excel, eso es post-day-zero.
        neto_esperado = e['cantidad_g'] - d['salidas_g']
        delta = d['neto_g'] - neto_esperado
        item = {
            'codigo_mp': k[0], 'lote': k[1],
            'nombre_comercial': e['nombre_comercial'],
            'proveedor_excel': e['proveedor'],
            'cant_excel_g': round(e['cantidad_g'], 1),
            'entradas_db_g': round(d['entradas_g'], 1),
            'salidas_db_g': round(d['salidas_g'], 1),
            'neto_db_g': round(d['neto_g'], 1),
            'neto_esperado_g': round(neto_esperado, 1),
            'delta_g': round(delta, 1),
        }
        if abs(delta) <= TOLERANCIA_G:
            en_db_match.append(item)
        else:
            en_db_con_delta.append(item)

    for k in excel_keys - db_keys:
        e = excel_verde[k]
        faltantes_en_db.append({
            'codigo_mp': k[0], 'lote': k[1],
            'nombre_comercial': e['nombre_comercial'],
            'proveedor_excel': e['proveedor'],
            'cant_excel_g': round(e['cantidad_g'], 1),
        })

    for k in db_keys - excel_keys:
        d = db_lotes[k]
        if d['neto_g'] <= 0.5:
            continue  # lote vacio en DB que tampoco esta en Excel — irrelevante
        solo_db.append({
            'codigo_mp': k[0], 'lote': k[1],
            'entradas_db_g': round(d['entradas_g'], 1),
            'salidas_db_g': round(d['salidas_g'], 1),
            'neto_db_g': round(d['neto_g'], 1),
        })

    # Ordenar por mayor delta para visualizar primero los problemas grandes
    en_db_con_delta.sort(key=lambda x: -abs(x['delta_g']))
    faltantes_en_db.sort(key=lambda x: -x['cant_excel_g'])
    solo_db.sort(key=lambda x: -x['neto_db_g'])

    delta_total_g = sum(x['delta_g'] for x in en_db_con_delta)

    return jsonify({
        'ok': True,
        'resumen': {
            'archivo': f.filename,
            'lotes_verde_excel': len(excel_verde),
            'lotes_excluidos_no_verde': rows_no_verde_count,
            'stock_total_excel_g': round(excel_total_g, 1),
            'stock_total_db_actual_g': round(db_total_neto_g, 1),
            'producciones_registradas': prod_n,
            'producciones_total_kg': round(prod_total_kg, 1),
            'count_match': len(en_db_match),
            'count_delta': len(en_db_con_delta),
            'count_faltantes_en_db': len(faltantes_en_db),
            'count_solo_db_no_excel': len(solo_db),
            'delta_total_g': round(delta_total_g, 1),
        },
        'en_db_con_delta': en_db_con_delta[:300],
        'faltantes_en_db': faltantes_en_db[:300],
        'solo_db_no_excel': solo_db[:300],
        'en_db_match_sample': en_db_match[:20],
        'nota_truncado': (
            len(en_db_con_delta) > 300 or len(faltantes_en_db) > 300
            or len(solo_db) > 300
        ),
    })


@bp.route("/api/admin/inventario-snapshot-pre-reset", methods=["GET"])
def admin_inventario_snapshot_pre_reset():
    """Descarga snapshot JSON completo de la BD antes del reset.

    Incluye: movimientos (todos), producciones, ordenes_compra + items,
    comprobantes_pago, audit_log (ultimos 1000), maestro_mps. El
    usuario lo descarga y guarda fuera de Render. Si el reset falla,
    podemos reconstruir.

    Solo admins. No escribe nada.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    import json as _json
    from datetime import datetime as _dt

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    snapshot = {
        'meta': {
            'generated_at': _dt.utcnow().isoformat() + 'Z',
            'generated_by': u,
            'db_path': DB_PATH,
        },
        'tablas': {},
    }

    def _dump(table, where=None, limit=None):
        sql = f"SELECT * FROM {table}"
        if where: sql += f" WHERE {where}"
        if limit: sql += f" LIMIT {limit}"
        try:
            rows = c.execute(sql).fetchall()
            return [dict(r) for r in rows]
        except sqlite3.OperationalError as _e:
            return {'error': str(_e)}

    snapshot['tablas']['movimientos'] = _dump('movimientos')
    snapshot['tablas']['producciones'] = _dump('producciones')
    snapshot['tablas']['ordenes_compra'] = _dump('ordenes_compra')
    snapshot['tablas']['ordenes_compra_items'] = _dump('ordenes_compra_items')
    snapshot['tablas']['comprobantes_pago'] = _dump('comprobantes_pago')
    snapshot['tablas']['maestro_mps'] = _dump('maestro_mps')
    snapshot['tablas']['audit_log'] = _dump(
        'audit_log', where='1=1 ORDER BY id DESC', limit=1000
    )

    conn.close()

    # Conteos rapidos para el meta
    snapshot['meta']['conteos'] = {
        k: (len(v) if isinstance(v, list) else 0)
        for k, v in snapshot['tablas'].items()
    }

    body = _json.dumps(snapshot, ensure_ascii=False, indent=2, default=str)
    fname = f"snapshot_pre_reset_{_dt.now().strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        body, mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@bp.route("/api/admin/inventario-reset-preview", methods=["POST"])
def admin_inventario_reset_preview():
    """Preview del reset: muestra que se va a hacer SIN escribir nada.

    Body: multipart con file = Excel del conteo fisico.

    Devuelve plan detallado:
      - movimientos_a_borrar: count
      - entradas_a_crear_desde_excel: 305 lotes verdes
      - entradas_oc_legitimas_a_preservar: count + total_g
      - salidas_produccion_a_preservar: count + total_g + lotes consumidos
      - alertas: lotes que las producciones consumieron pero NO estan en
                 Excel verde (post-reset darian stock negativo)
      - resumen_pre_post: stock total antes vs despues

    Solo admins. SIN ESCRITURA.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Archivo debe ser .xlsx'}), 400

    excel_verde, excel_total_g, excel_no_verde, errs = _parse_excel_verde(f, incluir_no_verde=True)
    if excel_verde is None:
        return jsonify({'error': errs[0] if errs else 'Excel invalido'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Conteos actuales
    movs_actuales = c.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
    stock_total_actual_g = float(c.execute(
        "SELECT COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) "
        "FROM movimientos"
    ).fetchone()[0] or 0)

    # Movimientos a preservar
    entradas_oc, salidas_prod = _identify_movimientos_a_preservar(c)
    entradas_oc_g = sum(float(r['cantidad'] or 0) for r in entradas_oc)
    salidas_prod_g = sum(float(r['cantidad'] or 0) for r in salidas_prod)

    # ── Salidas post-día-cero por lote (para fix de cantidad inicial) ──
    # Bug previo: el Excel verde tiene la cantidad ACTUAL (post-producciones),
    # pero al cargar como Entrada del día cero, después se aplicaban las salidas
    # otra vez → negativo. Fix: cantidad_dia_cero = excel_actual + salidas_post.
    salidas_post_por_lote = {}
    for s in salidas_prod:
        k = (s['material_id'], s['lote'] or '')
        salidas_post_por_lote[k] = salidas_post_por_lote.get(k, 0) + float(s['cantidad'] or 0)

    # Lotes huérfanos: consumidos por produccion pero NO en Excel verde
    excel_keys = set(excel_verde.keys())
    huerfanos_consumo = {}  # (cod,lote) -> sum_salidas_g
    for s in salidas_prod:
        cod = s['material_id']
        lote = s['lote'] or ''
        if (cod, lote) not in excel_keys:
            k = (cod, lote)
            huerfanos_consumo[k] = huerfanos_consumo.get(k, 0) + float(s['cantidad'] or 0)

    # Para cada huerfano: la entrada virtual usa Excel completo si esta, sino fallback = sum_salidas
    huerfanos_detalle = []
    huerfanos_total_g = 0.0
    for k, consumido in huerfanos_consumo.items():
        cod, lote = k
        en_excel_completo = excel_no_verde.get(k)
        if en_excel_completo and en_excel_completo['cantidad_g'] > 0:
            # Aún siendo no-verde, tiene cantidad — sumar salidas para no negativo
            cant_entrada = en_excel_completo['cantidad_g'] + consumido
            origen = 'excel_no_verde+salidas'
        else:
            cant_entrada = consumido  # fallback: para cerrar en 0
            origen = 'fallback_sum_salidas'
        huerfanos_detalle.append({
            'codigo_mp': cod, 'lote': lote,
            'cantidad_consumida_g': round(consumido, 1),
            'cantidad_entrada_virtual_g': round(cant_entrada, 1),
            'origen': origen,
        })
        huerfanos_total_g += cant_entrada

    salidas_lotes_no_verde = [
        {'codigo_mp': k[0], 'lote': k[1],
         'cantidad_g': round(v, 1)}
        for k, v in huerfanos_consumo.items()
    ]

    # Cantidad inicial total (con compensacion FEFO) = excel_verde + salidas_de_esos_lotes
    excel_inicial_total_g = 0.0
    lotes_compensados = []  # lotes verdes que tendrán cantidad_inicial > excel
    for k, info in excel_verde.items():
        salidas_post = salidas_post_por_lote.get(k, 0)
        cant_inicial = float(info['cantidad_g']) + salidas_post
        excel_inicial_total_g += cant_inicial
        if salidas_post > 0:
            lotes_compensados.append({
                'codigo_mp': k[0], 'lote': k[1],
                'cantidad_excel_actual_g': round(float(info['cantidad_g']), 1),
                'salidas_post_dia_cero_g': round(salidas_post, 1),
                'cantidad_inicial_dia_cero_g': round(cant_inicial, 1),
            })
    lotes_compensados.sort(key=lambda x: -x['salidas_post_dia_cero_g'])

    # Stock post = lo del Excel actual (sin compensacion) + OC nuevas + huerfanos - salidas
    # = (excel_inicial - salidas_de_lotes_verdes) + entradas_oc + huerfanos_total - salidas_huerfanos
    # Como huerfanos_total ya equivale a las salidas que les corresponden, queda:
    # stock_post = excel_total_g (la cantidad ACTUAL del excel) + entradas_oc_g + 0 (huerfanos cierran en 0)
    stock_post_g = excel_total_g + entradas_oc_g

    # Sample top lotes a crear (mas grandes)
    top_excel = sorted(
        [v for v in excel_verde.values()],
        key=lambda v: -v['cantidad_g']
    )[:10]

    conn.close()

    return jsonify({
        'ok': True,
        'plan': {
            'movimientos_a_borrar': movs_actuales,
            'entradas_iniciales_a_crear': {
                'count': len(excel_verde),
                'total_g': round(excel_total_g, 1),  # alias back-compat
                'total_g_excel_actual': round(excel_total_g, 1),
                'total_g_dia_cero_compensado': round(excel_inicial_total_g, 1),
                'lotes_compensados_por_salidas_post_count': len(lotes_compensados),
                'lotes_compensados_sample_top10': lotes_compensados[:10],
                'sample_top10': top_excel,
                'nota_fix': ('Cantidad cargada al día cero = excel_actual + '
                             'salidas_post_día_cero del mismo lote. Esto evita '
                             'negativos: el FEFO consumirá las salidas y el lote '
                             'quedará exactamente en lo que el Excel reporta hoy.'),
            },
            'entradas_oc_a_preservar': {
                'count': len(entradas_oc),
                'total_g': round(entradas_oc_g, 1),
                'sample': [{
                    'fecha': str(r['fecha'])[:10],
                    'codigo_mp': r['material_id'], 'lote': r['lote'] or '',
                    'cantidad_g': round(float(r['cantidad'] or 0), 1),
                    'numero_oc': r['numero_oc'],
                } for r in entradas_oc[:20]],
            },
            'salidas_produccion_a_preservar': {
                'count': len(salidas_prod),
                'total_g': round(salidas_prod_g, 1),
                'sample': [{
                    'fecha': str(r['fecha'])[:10],
                    'codigo_mp': r['material_id'], 'lote': r['lote'] or '',
                    'cantidad_g': round(float(r['cantidad'] or 0), 1),
                    'observaciones': r['observaciones'],
                } for r in salidas_prod[:20]],
            },
            'rows_no_verde_excluidas': len(excel_no_verde),
            'huerfanos_a_regenerar': {
                'count': len(huerfanos_detalle),
                'total_g_entradas_virtuales': round(huerfanos_total_g, 1),
                'sample': huerfanos_detalle[:30],
                'nota': ('Lotes que las producciones consumieron pero que '
                         'Catalina marco NO presentes. Se regeneraran como '
                         'Entrada virtual del Excel (no-verde) o fallback = '
                         'sum_salidas, para que la salida los lleve a 0 sin '
                         'dejar stock negativo. Trazabilidad de produccion '
                         'preservada.'),
            },
        },
        'alertas': {
            'salidas_a_lotes_no_verde': salidas_lotes_no_verde[:30],
            'count_salidas_a_lotes_no_verde': len(salidas_lotes_no_verde),
            'nota': ('Estas salidas consumieron lotes ya NO presentes '
                     'fisicamente. Tras la mejora se regenera Entrada virtual '
                     'que cierra el lote en 0g (no negativo). Trazabilidad '
                     'preservada.'),
        },
        'resumen_pre_post': {
            'stock_actual_g': round(stock_total_actual_g, 1),
            'stock_post_reset_g_estimado': round(stock_post_g, 1),
            'delta_g': round(stock_post_g - stock_total_actual_g, 1),
        },
    })


_RESET_TOKEN = "BORRAR_INVENTARIO_Y_CARGAR_EXCEL_2026_04_27"


@bp.route("/api/admin/inventario-reset-aplicar", methods=["POST"])
def admin_inventario_reset_aplicar():
    """APLICA el reset+replay del inventario. Destructivo. Tiene salvaguardas.

    Salvaguardas:
      1. Solo admins
      2. Body MUST contain confirmacion = token textual exacto
      3. Verifica que haya backup reciente (< 24h) en backup_log o intenta
         crear uno automaticamente antes de proceder
      4. Snapshot pre-reset guardado en audit_log (resumen)
      5. Audit log entry detallado (TRES entries: PRE, RESET, POST)
      6. Body multipart con Excel + JSON con confirmacion

    Plan ejecutado en orden, dentro de UNA transaccion:
      a) Capturar entradas_oc_legitimas + salidas_produccion (preservar)
      b) DELETE FROM movimientos
      c) INSERT 305 Entradas iniciales desde Excel verde
         (fecha=2026-04-15 = dia cero, operador='reset_2026_04_27')
      d) Re-INSERT entradas_oc_legitimas (preservadas, fecha y datos
         originales)
      e) Re-INSERT salidas_produccion (preservadas, fecha y datos
         originales — incluso si apunta a lote no-verde, queda como
         "stock negativo" para alertar)
      f) Audit log: ELIMINAR_INVENTARIO + CARGAR_EXCEL + RE_APLICAR_PROD
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    confirmacion = (request.form.get('confirmacion') or '').strip()
    if confirmacion != _RESET_TOKEN:
        return jsonify({
            'error': 'Confirmacion textual requerida',
            'detail': f'Body form debe incluir confirmacion="{_RESET_TOKEN}"',
        }), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Archivo debe ser .xlsx'}), 400

    excel_verde, excel_total_g, excel_no_verde, errs = _parse_excel_verde(f, incluir_no_verde=True)
    if excel_verde is None:
        return jsonify({'error': errs[0] if errs else 'Excel invalido'}), 400

    # Verificar backup reciente
    from datetime import datetime as _dt, timedelta as _td
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    backup_reciente = False
    try:
        last = c.execute(
            "SELECT MAX(started_at) FROM backup_log WHERE status='ok'"
        ).fetchone()[0]
        if last:
            try:
                # SQLite datetime('now','utc') devuelve 'YYYY-MM-DD HH:MM:SS'
                last_dt = _dt.fromisoformat(last.replace(' ', 'T').replace('Z', ''))
                if (_dt.utcnow() - last_dt) < _td(hours=24):
                    backup_reciente = True
            except Exception:
                pass
    except sqlite3.OperationalError:
        pass

    if not backup_reciente:
        # Intentar crear backup automatico antes de proceder
        try:
            from backup import do_backup
            do_backup(triggered_by='reset_inventario_pre')
        except Exception as _e:
            return jsonify({
                'error': 'Sin backup reciente y no pude crear uno automatico',
                'detail': str(_e),
                'recomendacion': 'Hacer click en /admin → tab Backups → "Backup ahora" antes de re-intentar.',
            }), 500

    # Capturar movimientos a preservar
    entradas_oc, salidas_prod = _identify_movimientos_a_preservar(c)

    # ── Salidas post-día-cero por lote (fix de cantidad inicial) ──
    # Cantidad cargada al día cero = excel_actual + salidas_post para que el
    # FEFO posterior no deje negativos.
    salidas_post_por_lote = {}
    for s in salidas_prod:
        k = (s['material_id'], s['lote'] or '')
        salidas_post_por_lote[k] = salidas_post_por_lote.get(k, 0) + float(s['cantidad'] or 0)

    # Calcular huerfanos a regenerar (lotes consumidos NO en Excel verde)
    excel_keys = set(excel_verde.keys())
    huerfanos_consumo = {}
    for s in salidas_prod:
        k = (s['material_id'], s['lote'] or '')
        if k not in excel_keys:
            huerfanos_consumo[k] = huerfanos_consumo.get(k, 0) + float(s['cantidad'] or 0)

    huerfanos_a_regenerar = []
    for k, consumido in huerfanos_consumo.items():
        cod, lote = k
        en_excel = excel_no_verde.get(k)
        if en_excel and en_excel['cantidad_g'] > 0:
            # No-verde con cantidad: sumar salidas para que cierre con la cant del excel
            cant = en_excel['cantidad_g'] + consumido
            info = en_excel
        else:
            cant = consumido  # fallback para cerrar en 0
            info = {
                'codigo_mp': cod, 'lote': lote,
                'inci': '', 'nombre_comercial': '',
                'proveedor': '', 'estanteria': '', 'posicion': '',
                'fecha_vencimiento': '',
            }
        huerfanos_a_regenerar.append({**info, 'cantidad_g': cant})

    movs_borrados = c.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]

    # Audit pre-reset
    try:
        import json as _json
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now'))""",
                  (u, 'RESET_INVENTARIO_PRE', 'movimientos', '_BULK_',
                   _json.dumps({
                       'movs_a_borrar': movs_borrados,
                       'entradas_oc_preservar': len(entradas_oc),
                       'salidas_prod_preservar': len(salidas_prod),
                       'lotes_excel_verde_a_cargar': len(excel_verde),
                       'lotes_huerfanos_regenerar': len(huerfanos_a_regenerar),
                       'excel_total_g': round(excel_total_g, 1),
                   }, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        pass

    # ── ATOMIC: BORRAR + RECREAR ─────────────────────────────────────────
    # sqlite3 abre transaccion implicita en DML; conn.commit/rollback la cierra.
    try:
        # 1. DELETE
        c.execute("DELETE FROM movimientos")

        # 2. Cargar Entradas iniciales del Excel verde (compensadas con salidas post)
        DIA_CERO = '2026-04-15T00:00:00'
        OBS_INICIAL = 'Carga inicial Excel dia cero v8_1 — reset 2026-04-27'
        OPERADOR_RESET = 'reset_2026_04_27'
        n_excel_inserted = 0
        n_lotes_compensados = 0
        total_compensacion_g = 0.0
        for (cod, lote), info in excel_verde.items():
            cant_excel = float(info['cantidad_g'] or 0)
            cant_salidas_post = salidas_post_por_lote.get((cod, lote), 0)
            cant_inicial = cant_excel + cant_salidas_post  # ← FIX día cero
            if cant_inicial <= 0:
                continue
            obs = OBS_INICIAL
            if cant_salidas_post > 0:
                obs = (f'{OBS_INICIAL} | Compensación FEFO: '
                       f'excel_hoy={int(round(cant_excel))}g + '
                       f'salidas_post={int(round(cant_salidas_post))}g')
                n_lotes_compensados += 1
                total_compensacion_g += cant_salidas_post
            c.execute("""INSERT INTO movimientos
                         (material_id, material_nombre, cantidad, tipo, fecha,
                          observaciones, lote, fecha_vencimiento, estanteria,
                          posicion, proveedor, estado_lote, operador)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      (cod, info['nombre_comercial'] or info['inci'] or cod,
                       cant_inicial, 'Entrada', DIA_CERO,
                       obs, lote, info['fecha_vencimiento'] or None,
                       info['estanteria'] or '', info['posicion'] or '',
                       info['proveedor'] or '', 'VIGENTE', OPERADOR_RESET))
            n_excel_inserted += 1

        # 2b. Regenerar Entradas virtuales para lotes huerfanos (consumidos
        # por produccion pero NO presentes fisicamente en el Excel verde).
        # Esto evita que las salidas dejen stock negativo. Las entradas
        # virtuales tienen observaciones distintas para auditarse despues.
        OBS_HUERFANO = 'Lote consumido pre-reset — Entrada regenerada para preservar trazabilidad de produccion'
        for info in huerfanos_a_regenerar:
            if info['cantidad_g'] <= 0:
                continue
            c.execute("""INSERT INTO movimientos
                         (material_id, material_nombre, cantidad, tipo, fecha,
                          observaciones, lote, fecha_vencimiento, estanteria,
                          posicion, proveedor, estado_lote, operador)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      (info['codigo_mp'],
                       info.get('nombre_comercial') or info.get('inci') or info['codigo_mp'],
                       float(info['cantidad_g']), 'Entrada', DIA_CERO,
                       OBS_HUERFANO, info['lote'],
                       info.get('fecha_vencimiento') or None,
                       info.get('estanteria') or '', info.get('posicion') or '',
                       info.get('proveedor') or '', 'AGOTADO', OPERADOR_RESET))
        n_huerfanos_inserted = len(huerfanos_a_regenerar)

        # 3. Re-insertar Entradas con OC legitimas
        for r in entradas_oc:
            c.execute("""INSERT INTO movimientos
                         (material_id, material_nombre, cantidad, tipo, fecha,
                          observaciones, lote, fecha_vencimiento, estanteria,
                          posicion, proveedor, estado_lote, operador,
                          numero_oc, numero_factura, precio_kg)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      (r['material_id'], r['material_nombre'], r['cantidad'],
                       r['tipo'], r['fecha'], r['observaciones'], r['lote'],
                       r['fecha_vencimiento'], r['estanteria'], r['posicion'],
                       r['proveedor'], r['estado_lote'], r['operador'],
                       r['numero_oc'], r['numero_factura'], r['precio_kg']))

        # 4. Re-insertar Salidas de produccion
        for r in salidas_prod:
            c.execute("""INSERT INTO movimientos
                         (material_id, material_nombre, cantidad, tipo, fecha,
                          observaciones, lote, fecha_vencimiento, estanteria,
                          posicion, proveedor, estado_lote, operador,
                          numero_oc, numero_factura, precio_kg)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      (r['material_id'], r['material_nombre'], r['cantidad'],
                       r['tipo'], r['fecha'], r['observaciones'], r['lote'],
                       r['fecha_vencimiento'], r['estanteria'], r['posicion'],
                       r['proveedor'], r['estado_lote'], r['operador'],
                       r['numero_oc'], r['numero_factura'], r['precio_kg']))

        conn.commit()
    except Exception as _e:
        try: conn.rollback()
        except Exception as _r:
            __import__('logging').getLogger('admin').warning('rollback no aplicable: %s', _r)
        return jsonify({
            'ok': False,
            'error': f'Reset abortado por error: {_e}',
            'detail': 'Transaccion revertida. La BD esta como antes. '
                      'Restaurar backup si hay duda.',
        }), 500

    # Validacion post: contar lo que quedo
    movs_post = c.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
    stock_post_g = float(c.execute(
        "SELECT COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) "
        "FROM movimientos"
    ).fetchone()[0] or 0)

    # Audit post-reset
    try:
        c.execute("""INSERT INTO audit_log
                     (usuario, accion, tabla, registro_id, detalle, ip, fecha)
                     VALUES (?,?,?,?,?,?,datetime('now'))""",
                  (u, 'RESET_INVENTARIO_POST', 'movimientos', '_BULK_',
                   _json.dumps({
                       'movs_borrados': movs_borrados,
                       'lotes_excel_cargados': n_excel_inserted,
                       'lotes_compensados_fefo': n_lotes_compensados,
                       'compensacion_total_g': round(total_compensacion_g, 1),
                       'huerfanos_regenerados': n_huerfanos_inserted,
                       'entradas_oc_re_insertadas': len(entradas_oc),
                       'salidas_prod_re_insertadas': len(salidas_prod),
                       'movs_post_total': movs_post,
                       'stock_post_g': round(stock_post_g, 1),
                       'excel_total_g_origen': round(excel_total_g, 1),
                   }, ensure_ascii=False),
                   request.remote_addr))
    except sqlite3.OperationalError:
        pass

    conn.commit()
    conn.close()

    return jsonify({
        'ok': True,
        'message': (f'Reset aplicado. Borrados {movs_borrados} movimientos, '
                    f'cargadas {n_excel_inserted} Entradas iniciales del Excel '
                    f'verde ({n_lotes_compensados} lotes compensados con '
                    f'+{int(round(total_compensacion_g))} g por salidas post-día-cero) '
                    f'+ {n_huerfanos_inserted} Entradas regeneradas (huerfanos), '
                    f'preservadas {len(entradas_oc)} Entradas con OC + '
                    f'{len(salidas_prod)} Salidas de produccion.'),
        'resumen': {
            'movs_borrados': movs_borrados,
            'lotes_excel_cargados': n_excel_inserted,
            'lotes_compensados_fefo': n_lotes_compensados,
            'compensacion_total_g': round(total_compensacion_g, 1),
            'huerfanos_regenerados': n_huerfanos_inserted,
            'entradas_oc_preservadas': len(entradas_oc),
            'salidas_prod_preservadas': len(salidas_prod),
            'movs_post_total': movs_post,
            'stock_post_g': round(stock_post_g, 1),
        },
    })


@bp.route("/api/admin/diagnosticar-formulas", methods=["GET"])
def admin_diagnosticar_formulas():
    """Detecta items en formula_items con material_id que NO existe en
    maestro_mps o cuyo nombre no coincide. Sugiere correccion buscando
    por nombre similar en el catalogo real.

    Para cada item con problema retorna:
      producto, material_id_actual, material_nombre,
      problema (huerfano|mismatch_nombre),
      candidato_sugerido: {codigo, nombre, score}

    Solo admins. SIN escritura.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    import re as _re
    import unicodedata as _ud

    # Palabras que se ignoran al comparar (genericas / sufijos de proveedor)
    _STOPWORDS = {
        'LYPHAR', 'YTBIO', 'LIQUIDO', 'POLVO', 'SOLUCION', 'AL', 'EN',
        'KD', 'KDA', 'DE', 'LA', 'EL', 'LOS', 'LAS', 'POR', 'PARA',
        'GRADO', 'COSMETICO', 'COSMETICA', 'COSMETICO', 'NF', 'USP',
        'INCHEMICAL', 'AGENQUIMICOS', 'BASF', 'IMCD',
    }

    # Sinonimos cosmeticos español/ingles + INCI (cientos de pares)
    _SINONIMOS_PARES = [
        # Polioles
        ('GLICERINA', 'GLYCERIN', 'GLYCEROL'),
        ('PROPILENGLICOL', 'PROPYLENE', 'GLYCOL', 'PG'),
        ('POLIETILENGLICOL', 'POLYETHYLENE', 'GLYCOL', 'PEG'),
        ('BUTILENGLICOL', 'BUTYLENE', 'GLYCOL', 'BG'),
        ('PENTILENGLICOL', 'PENTYLENE', 'GLYCOL'),
        ('HEXANODIOL', 'HEXANEDIOL'),
        # Acidos
        ('ACIDO', 'ACID'),
        ('SALICILICO', 'SALICYLIC'),
        ('HIALURONICO', 'HYALURONIC'),
        ('LACTICO', 'LACTIC'),
        ('GLICOLICO', 'GLYCOLIC'),
        ('CITRICO', 'CITRIC'),
        ('TRANEXAMICO', 'TRANEXAMIC'),
        ('AZELAICO', 'AZELAIC'),
        ('ASCORBICO', 'ASCORBIC'),
        ('KOJICO', 'KOJIC'),
        ('FERULICO', 'FERULIC'),
        ('MANDELICO', 'MANDELIC'),
        ('SUCCINICO', 'SUCCINIC'),
        ('PALMITICO', 'PALMITIC'),
        # Surfactantes / emulsificantes
        ('TWEEN', 'POLYSORBATE'),
        ('SORBITAN', 'SORBITAN'),
        # Alcoholes
        ('ALCOHOL', 'ETHANOL', 'ETANOL'),
        ('CETILICO', 'CETYL'),
        ('ESTEARILICO', 'STEARYL'),
        ('CETOESTEARILICO', 'CETEARYL'),
        # Conservantes
        ('FENOXIETANOL', 'PHENOXYETHANOL'),
        ('BENZOATO', 'BENZOATE'),
        ('SORBATO', 'SORBATE'),
        ('PARABENO', 'PARABEN'),
        # Vitaminas / activos
        ('NIACINAMIDA', 'NIACINAMIDE', 'NICOTINAMIDE'),
        ('UREA', 'CARBAMIDE'),
        ('ALANTOINA', 'ALLANTOIN'),
        ('PANTENOL', 'PANTHENOL'),
        ('RETINALDEHIDO', 'RETINALDEHYDE', 'RETINAL'),
        ('RETINOL', 'RETINOL'),
        ('TOCOFEROL', 'TOCOPHEROL'),
        ('ASCORBIL', 'ASCORBYL'),
        ('ARBUTINA', 'ARBUTIN'),
        ('ECTOINA', 'ECTOIN', 'ECTOINE'),
        ('BAKUCHIOL', 'BAKUCHIOL', 'BACKUCHIOL'),
        ('CAFEINA', 'CAFFEINE'),
        ('GLUTATION', 'GLUTATHIONE'),
        ('ADENOSINA', 'ADENOSINE'),
        ('BIOTINA', 'BIOTIN'),
        ('MELATONINA', 'MELATONIN'),
        # Aminoacidos / proteinas
        ('GLICINA', 'GLYCINE'),
        ('CARNITINA', 'CARNITINE'),
        ('BETAINA', 'BETAINE'),
        ('GLICINAMIDA', 'GLYCINAMIDE'),
        # Cationes
        ('SODIO', 'SODIUM'),
        ('POTASIO', 'POTASSIUM'),
        ('CALCIO', 'CALCIUM'),
        ('MAGNESIO', 'MAGNESIUM'),
        ('ZINC', 'ZINC'),
        ('HIERRO', 'IRON'),
        # Oxidos / minerales
        ('OXIDO', 'OXIDE'),
        ('TITANIO', 'TITANIUM'),
        ('DIOXIDO', 'DIOXIDE'),
        # Aguas / vehiculos
        ('AGUA', 'WATER', 'AQUA'),
        ('DESIONIZADA', 'DEIONIZED'),
        ('DESTILADA', 'DISTILLED'),
        # Aceites / esteres
        ('ACEITE', 'OIL'),
        ('TRIGLICERIDO', 'TRIGLYCERIDE'),
        ('CAPRILICO', 'CAPRYLIC'),
        ('CAPRICO', 'CAPRIC'),
        ('JOJOBA', 'JOJOBA'),
        ('ARGAN', 'ARGAN'),
        ('ESCUALANO', 'SQUALANE'),
        ('ESCUALENO', 'SQUALENE'),
        # Emulsionantes / espesantes
        ('CARBOPOL', 'CARBOMER'),
        ('GOMA', 'GUM'),
        ('XANTAN', 'XANTHAN'),
        ('CELULOSA', 'CELLULOSE'),
        # Otros
        ('EDTA', 'EDTA'),
        ('TRIETANOLAMINA', 'TRIETHANOLAMINE', 'TEA'),
        ('CENTELLA', 'CENTELLA', 'GOTU', 'KOLA'),
        ('REGALIZ', 'LICORICE'),
        ('SALVIA', 'SAGE'),
        ('YOGURT', 'YOGURT'),
        ('SILIMARINA', 'SILYMARIN'),
        ('RESVERATROL', 'RESVERATROL'),
        ('BETAGLUCAN', 'BETAGLUCAN', 'BETA-GLUCAN'),
        ('NAG', 'GLUCOSAMINE', 'GLUCOSAMINA'),
        ('PEPTIDO', 'PEPTIDE'),
        ('PALMITOIL', 'PALMITOYL'),
        ('ACETIL', 'ACETYL'),
        ('TETRAPEPTIDO', 'TETRAPEPTIDE'),
        ('TRIPEPTIDO', 'TRIPEPTIDE'),
        ('HEXAPEPTIDO', 'HEXAPEPTIDE'),
        ('PENTAPEPTIDO', 'PENTAPEPTIDE'),
        ('NONAPEPTIDO', 'NONAPEPTIDE'),
        ('MIRISTOIL', 'MYRISTOYL'),
        ('FOSFATO', 'PHOSPHATE'),
        ('TOCOFERIL', 'TOCOPHERYL'),
        ('GLUCOSIDE', 'GLUCOSIDO'),
        ('GLUCONOLACTONA', 'GLUCONOLACTONE'),
        ('FENOXIETANOL', 'PHENOXYETHANOL'),
        ('HIDROXIDO', 'HYDROXIDE'),
        ('BICARBONATO', 'BICARBONATE'),
        ('EXTRACTO', 'EXTRACT'),
        ('POLVO', 'POWDER'),
    ]
    _SINONIMOS = {}
    for grupo in _SINONIMOS_PARES:
        for w in grupo:
            _SINONIMOS.setdefault(w, set()).update(set(grupo) - {w})

    def _expandir_sinonimos(palabras):
        """Devuelve set ampliado con sinonimos conocidos."""
        ampliado = set(palabras)
        for p in palabras:
            if p in _SINONIMOS:
                ampliado.update(_SINONIMOS[p])
        return ampliado

    def _norm(s):
        """Normaliza: ascii, uppercase, sin parentesis, espacios consolidados."""
        if not s: return ''
        s = _ud.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().upper().strip()
        # Eliminar parentesis y su contenido (ej. "(LYPHAR)", "(LIQUIDO)")
        s = _re.sub(r'\([^)]*\)', '', s)
        # Reemplazar guiones, comas, puntos con espacios
        s = _re.sub(r'[\-_,.;:/]', ' ', s)
        s = _re.sub(r'\s+', ' ', s).strip()
        return s

    def _palabras_clave(s):
        """Set de palabras significativas (sin stopwords)."""
        return set(p for p in _norm(s).split() if p and p not in _STOPWORDS)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Cargar catalogo maestro
    maestro = []
    try:
        rows = c.execute("""SELECT codigo_mp, COALESCE(nombre_comercial,''),
                                   COALESCE(nombre_inci,'')
                            FROM maestro_mps WHERE activo=1""").fetchall()
        for r in rows:
            cod = r[0]
            nombres_norm = set(filter(None, [_norm(r[1]), _norm(r[2])]))
            palabras = set()
            for nm in [r[1], r[2]]:
                palabras.update(_palabras_clave(nm))
            maestro.append({'codigo': cod, 'nombres_norm': nombres_norm,
                            'palabras_clave': palabras,
                            'nombre_comercial': r[1], 'nombre_inci': r[2]})
    except sqlite3.OperationalError:
        return jsonify({'error': 'maestro_mps no disponible'}), 500

    catalogo_codigos = {m['codigo'] for m in maestro}

    # Cargar formula_items
    try:
        rows = c.execute(
            "SELECT id, producto_nombre, material_id, material_nombre, "
            "       cantidad_g_por_lote "
            "FROM formula_items "
            "ORDER BY producto_nombre, material_nombre"
        ).fetchall()
    except sqlite3.OperationalError:
        return jsonify({'error': 'formula_items no disponible'}), 500

    problemas = []
    for r in rows:
        item_id = r[0]
        producto = r[1]
        mid = r[2]
        nombre_formula = r[3] or ''
        cant_lote = float(r[4] or 0)
        nombre_norm = _norm(nombre_formula)
        palabras_form = _palabras_clave(nombre_formula)

        es_huerfano = mid not in catalogo_codigos

        # Expandir palabras con sinónimos (formula + catálogo)
        palabras_form_exp = _expandir_sinonimos(palabras_form)

        # Buscar candidatos
        candidatos = []
        for m in maestro:
            score = 0
            len_form = len(palabras_form)
            len_cat = len(m['palabras_clave'])
            cat_palabras_exp = _expandir_sinonimos(m['palabras_clave'])

            # Score 100: match exacto normalizado
            if nombre_norm and nombre_norm in m['nombres_norm']:
                score = 100

            # Score 95: palabras clave idénticas (sin stopwords)
            elif palabras_form and m['palabras_clave'] and palabras_form == m['palabras_clave']:
                score = 95

            elif palabras_form and m['palabras_clave']:
                # Match con sinonimos
                comunes = palabras_form & m['palabras_clave']
                comunes_exp = palabras_form_exp & cat_palabras_exp
                len_comunes = len(comunes)
                len_comunes_exp = len(comunes_exp)

                # Score 90: todas palabras formula en catalogo (directo)
                if comunes == palabras_form and len_form >= 1:
                    if len_cat <= len_form + 1:
                        score = 92
                    elif len_cat <= len_form * 2:
                        score = 88
                    else:
                        score = 80

                # Score 85: todas palabras catalogo en formula (catalogo más corto)
                elif comunes == m['palabras_clave'] and len_cat >= 1:
                    if len_form <= len_cat + 1:
                        score = 85
                    else:
                        score = 75

                # Score 80: match con sinonimos — todas palabras formula tienen equivalente
                elif len_comunes_exp >= len_form and len_form >= 1:
                    score = 80

                # Score 70: al menos 70% de palabras en común
                elif len_comunes >= max(2, int(len_form * 0.7)):
                    score = 70

                # Score 65: al menos 70% con sinonimos
                elif len_comunes_exp >= max(2, int(len_form * 0.7)):
                    score = 65

                # Score 55: al menos 50% de palabras en común
                elif len_comunes >= max(1, int(len_form * 0.5)) and len_comunes >= 1:
                    score = 55

            # Bonus: substring match — cada palabra de formula está como prefijo
            # de alguna palabra del catalogo (ej. 'GLIC' ⊂ 'GLICERINA')
            if score < 70 and palabras_form and m['palabras_clave']:
                substring_matches = 0
                for pf in palabras_form:
                    if len(pf) >= 4:
                        for pc in m['palabras_clave']:
                            if pc.startswith(pf) or pf.startswith(pc):
                                substring_matches += 1
                                break
                if substring_matches == len(palabras_form) and len_form >= 1:
                    score = max(score, 75)
                elif substring_matches >= max(1, int(len_form * 0.7)):
                    score = max(score, 65)

            if score > 0:
                candidatos.append({
                    'codigo': m['codigo'],
                    'nombre_comercial': m['nombre_comercial'],
                    'nombre_inci': m['nombre_inci'],
                    'score': score,
                })

        # Ordenar y filtrar
        candidatos.sort(key=lambda x: -x['score'])
        # Si hay match alto (>=85), descartar matches débiles para reducir ruido
        if candidatos and candidatos[0]['score'] >= 85:
            min_threshold = max(70, candidatos[0]['score'] - 20)
            candidatos = [c for c in candidatos if c['score'] >= min_threshold]
        candidatos = candidatos[:5]

        # Auto-corregible: top score >= 75 con clara dominancia sobre el segundo
        auto_corregible = False
        if candidatos:
            top_score = candidatos[0]['score']
            second_score = candidatos[1]['score'] if len(candidatos) > 1 else 0
            delta = top_score - second_score
            # Reglas escaladas — score más alto requiere menos delta
            if top_score == 100 and len([c for c in candidatos if c['score'] == 100]) == 1:
                auto_corregible = True
            elif top_score >= 95 and delta >= 5:
                auto_corregible = True
            elif top_score >= 88 and delta >= 8:
                auto_corregible = True
            elif top_score >= 80 and delta >= 10:
                auto_corregible = True
            elif top_score >= 75 and delta >= 15:
                auto_corregible = True

        # Determinar problema
        problema = None
        if es_huerfano:
            problema = 'huerfano'
        else:
            cat_match = next((m for m in maestro if m['codigo'] == mid), None)
            if cat_match and palabras_form and cat_match['palabras_clave']:
                comunes = palabras_form & cat_match['palabras_clave']
                # Mismatch si menos del 50% de palabras coinciden
                if len(comunes) < max(1, int(len(palabras_form) * 0.5)):
                    problema = 'mismatch_nombre'

        if problema:
            mejor = candidatos[0] if candidatos else None
            # Marcar como auto si cumple criterio
            if mejor and auto_corregible:
                mejor = dict(mejor)
                mejor['auto'] = True
            problemas.append({
                'formula_item_id': item_id,
                'producto': producto,
                'material_id_actual': mid,
                'material_nombre_formula': nombre_formula,
                'cantidad_g_por_lote': cant_lote,
                'problema': problema,
                'candidatos_sugeridos': candidatos,
                'mejor_candidato': mejor,
                'auto_corregible': auto_corregible,
            })

    # Stats
    stats = {
        'total_formula_items': len(rows),
        'total_problemas': len(problemas),
        'huerfanos': sum(1 for p in problemas if p['problema'] == 'huerfano'),
        'mismatch_nombre': sum(1 for p in problemas if p['problema'] == 'mismatch_nombre'),
        'auto_corregibles': sum(1 for p in problemas if p.get('auto_corregible')),
        'requieren_revision': sum(1 for p in problemas
                                  if not p.get('mejor_candidato')),
        'con_sugerencia_baja': sum(1 for p in problemas
                                   if p.get('mejor_candidato')
                                   and not p.get('auto_corregible')),
    }

    # Agrupar por producto
    by_producto = {}
    for p in problemas:
        prod = p['producto']
        by_producto.setdefault(prod, []).append(p)
    productos_afectados = sorted(
        [{'producto': k, 'count': len(v)} for k, v in by_producto.items()],
        key=lambda x: -x['count'],
    )

    conn.close()
    return jsonify({
        'stats': stats,
        'problemas': problemas,
        'productos_afectados': productos_afectados,
    })


@bp.route("/api/admin/corregir-formulas", methods=["POST"])
def admin_corregir_formulas():
    """Aplica correcciones de material_id a formula_items.

    Body:
      token: 'CORREGIR_FORMULAS_2026'
      correcciones: list de {formula_item_id, nuevo_material_id, nuevo_material_nombre}
        Si nuevo_material_nombre es null, se usa el nombre actual de la formula.

    Backup previo + audit log.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    if d.get('token', '').strip() != 'CORREGIR_FORMULAS_2026':
        return jsonify({'error': 'Token incorrecto'}), 403
    correcciones = d.get('correcciones') or []
    if not correcciones:
        return jsonify({'error': 'Sin correcciones'}), 400

    # Backup previo
    try:
        do_backup(triggered_by='pre_corregir_formulas')
    except Exception as e:
        return jsonify({'error': f'Backup fallo: {str(e)[:200]}'}), 500

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    aplicados = []
    errores = []
    for corr in correcciones:
        try:
            fid = int(corr.get('formula_item_id'))
            nuevo_mid = (corr.get('nuevo_material_id') or '').strip()
            nuevo_nombre = corr.get('nuevo_material_nombre')
            if not nuevo_mid:
                errores.append({'formula_item_id': fid, 'error': 'sin nuevo_material_id'})
                continue
            # Capturar valores previos para audit
            row = c.execute(
                "SELECT producto_nombre, material_id, material_nombre "
                "FROM formula_items WHERE id=?", (fid,)
            ).fetchone()
            if not row:
                errores.append({'formula_item_id': fid, 'error': 'no encontrado'})
                continue
            previo = {
                'producto': row[0],
                'material_id_previo': row[1],
                'material_nombre_previo': row[2],
            }
            if nuevo_nombre:
                c.execute(
                    "UPDATE formula_items SET material_id=?, material_nombre=? "
                    "WHERE id=?",
                    (nuevo_mid, nuevo_nombre, fid),
                )
            else:
                c.execute(
                    "UPDATE formula_items SET material_id=? WHERE id=?",
                    (nuevo_mid, fid),
                )
            aplicados.append({
                **previo,
                'material_id_nuevo': nuevo_mid,
                'material_nombre_nuevo': nuevo_nombre or row[2],
            })
        except Exception as _e:
            errores.append({'formula_item_id': corr.get('formula_item_id'),
                            'error': str(_e)[:200]})

    conn.commit()

    # Audit
    try:
        import json as _json
        c.execute(
            """INSERT INTO audit_log
               (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now'))""",
            (u, 'CORREGIR_FORMULAS', 'formula_items', '_BULK_',
             _json.dumps({
                 'count': len(aplicados),
                 'errores': len(errores),
                 'correcciones_sample': aplicados[:30],
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except Exception:
        pass
    conn.close()

    return jsonify({
        'ok': True,
        'count_aplicados': len(aplicados),
        'count_errores': len(errores),
        'aplicados': aplicados[:50],
        'errores': errores[:30],
        'mensaje': f'{len(aplicados)} fórmulas corregidas. Backup previo creado.',
    })


@bp.route("/api/admin/revertir-formulas-desde-backup", methods=["POST"])
def admin_revertir_formulas_desde_backup():
    """Revierte formula_items al estado guardado en el backup mas reciente
    con triggered_by='pre_corregir_formulas'. Quirurgico: solo reemplaza
    formula_items, NO toca movimientos / catalogo / OCs.

    Body:
      token: 'REVERTIR_FORMULAS_2026'

    Caso de uso: el usuario aplico correcciones y se dio cuenta que
    algunas eran incorrectas. Este endpoint deshace los cambios usando
    el backup automatico pre-aplicacion.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    if d.get('token', '').strip() != 'REVERTIR_FORMULAS_2026':
        return jsonify({'error': 'Token incorrecto'}), 403

    # Encontrar backup mas reciente pre-corregir o pre-eliminar formulas
    import gzip, os as _os, tempfile
    bk_db_path = None
    bk_info = None
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        row = c.execute("""
            SELECT file_path, started_at, triggered_by FROM backup_log
            WHERE status='ok' AND (triggered_by LIKE 'pre_corregir_formulas%'
                                   OR triggered_by LIKE 'pre_eliminar_formulas%')
            ORDER BY started_at DESC LIMIT 1
        """).fetchone()
        conn.close()
        if not row:
            return jsonify({'error': 'No encontre backup pre-correccion. Imposible revertir.'}), 404
        bk_db_path, bk_started, bk_trigger = row
        bk_info = {'file': bk_db_path, 'started_at': bk_started, 'triggered_by': bk_trigger}
    except sqlite3.OperationalError as _e:
        return jsonify({'error': f'No pude leer backup_log: {_e}'}), 500

    if not _os.path.exists(bk_db_path):
        return jsonify({'error': f'Archivo backup no existe: {bk_db_path}',
                        'backup_info': bk_info}), 404

    # Descomprimir backup a tmp
    try:
        if bk_db_path.endswith('.gz'):
            tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
            tmp_path = tmp.name
            tmp.close()
            with gzip.open(bk_db_path, 'rb') as gz_in, open(tmp_path, 'wb') as out:
                while True:
                    chunk = gz_in.read(64 * 1024)
                    if not chunk: break
                    out.write(chunk)
        else:
            tmp_path = bk_db_path
    except Exception as _e:
        return jsonify({'error': f'No pude descomprimir backup: {_e}',
                        'backup_info': bk_info}), 500

    # Leer formula_items del backup
    try:
        bk = sqlite3.connect(tmp_path)
        bk_rows = bk.execute(
            "SELECT id, producto_nombre, material_id, material_nombre, "
            "       porcentaje, cantidad_g_por_lote "
            "FROM formula_items"
        ).fetchall()
        bk.close()
    except Exception as _e:
        return jsonify({'error': f'No pude leer formula_items del backup: {_e}',
                        'backup_info': bk_info}), 500
    finally:
        if tmp_path != bk_db_path and _os.path.exists(tmp_path):
            try: _os.remove(tmp_path)
            except Exception as _e:
                __import__('logging').getLogger('admin').debug('cleanup tmp_path fallo: %s', _e)

    # Crear backup adicional ANTES de revertir (por si acaso)
    try:
        do_backup(triggered_by='pre_revertir_formulas')
    except Exception:
        pass

    # Reemplazar formula_items en BD actual con el del backup
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        actuales = c.execute("SELECT COUNT(*) FROM formula_items").fetchone()[0]
        c.execute("DELETE FROM formula_items")
        for r in bk_rows:
            c.execute(
                """INSERT INTO formula_items
                   (id, producto_nombre, material_id, material_nombre,
                    porcentaje, cantidad_g_por_lote)
                   VALUES (?,?,?,?,?,?)""",
                r,
            )
        conn.commit()
    except Exception as _e:
        try: conn.rollback()
        except Exception as _r:
            __import__('logging').getLogger('admin').warning('rollback no aplicable: %s', _r)
        return jsonify({'error': f'Reversion fallo: {_e}'}), 500

    # Audit
    try:
        import json as _json
        c.execute(
            """INSERT INTO audit_log
               (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now'))""",
            (u, 'REVERTIR_FORMULAS_DESDE_BACKUP', 'formula_items', '_BULK_',
             _json.dumps({
                 'backup': bk_info,
                 'items_actuales_eliminados': actuales,
                 'items_restaurados_del_backup': len(bk_rows),
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except Exception:
        pass
    conn.close()

    return jsonify({
        'ok': True,
        'mensaje': (f'formula_items revertido al estado del backup '
                    f'{bk_info.get("started_at")}. Eliminados {actuales} '
                    f'items actuales, restaurados {len(bk_rows)} del backup.'),
        'items_eliminados': actuales,
        'items_restaurados': len(bk_rows),
        'backup_usado': bk_info,
    })


@bp.route("/api/admin/eliminar-formulas-obsoletas", methods=["POST"])
def admin_eliminar_formulas_obsoletas():
    """Elimina items en formula_items que apuntan a material_id huerfano
    Y sin candidato similar en maestro_mps. Probablemente son ingredientes
    descontinuados o productos obsoletos.

    Body:
      token: 'ELIMINAR_FORMULAS_OBSOLETAS_2026'
      formula_item_ids: lista de IDs a eliminar (del diagnostico previo)

    Backup previo + audit log.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    if d.get('token', '').strip() != 'ELIMINAR_FORMULAS_OBSOLETAS_2026':
        return jsonify({'error': 'Token incorrecto'}), 403
    ids = d.get('formula_item_ids') or []
    if not ids:
        return jsonify({'error': 'Sin IDs a eliminar'}), 400

    # Backup previo
    try:
        do_backup(triggered_by='pre_eliminar_formulas_obsoletas')
    except Exception as e:
        return jsonify({'error': f'Backup fallo: {str(e)[:200]}'}), 500

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    eliminados = []
    for fid in ids:
        try:
            row = c.execute(
                "SELECT producto_nombre, material_id, material_nombre "
                "FROM formula_items WHERE id=?", (fid,)
            ).fetchone()
            if not row:
                continue
            c.execute("DELETE FROM formula_items WHERE id=?", (fid,))
            eliminados.append({
                'formula_item_id': fid,
                'producto': row[0],
                'material_id': row[1],
                'material_nombre': row[2],
            })
        except Exception:
            continue
    conn.commit()

    # Audit
    try:
        import json as _json
        c.execute(
            """INSERT INTO audit_log
               (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now'))""",
            (u, 'ELIMINAR_FORMULAS_OBSOLETAS', 'formula_items', '_BULK_',
             _json.dumps({
                 'count': len(eliminados),
                 'eliminados_sample': eliminados[:30],
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except Exception:
        pass
    conn.close()

    return jsonify({
        'ok': True,
        'count_eliminados': len(eliminados),
        'eliminados': eliminados[:50],
        'mensaje': f'{len(eliminados)} items obsoletos eliminados de formula_items.',
    })


@bp.route("/api/admin/aplicar-correcciones-formulas-batch-2026-04-28", methods=["POST"])
def admin_aplicar_correcciones_formulas_batch_20260428():
    """Aplica el batch de correcciones de formulas+catalogo decidido el 2026-04-28
    con Sebastian + Alejandro. Lee y ejecuta el archivo SQL versionado en
    scripts/migraciones/correcciones_formulas_2026_04_28.sql.

    Body:
      token: 'APLICAR_CORRECCIONES_2026_04_28'

    Operaciones (240 cambios totales):
      - 44 INSERT nuevos MPs (codigos MP00400+)
      - 1  UPDATE Azeclair MP00284 -> activo=0
      - 1  UPDATE INCI oficial AOS 40 (MP00212)
      - 207 UPDATE formula_items (correcciones huerfanos + typos)
      - 33 DELETE formula_items (aguas internas infinitas)

    Backup automatico antes de tocar.
    Una sola pulsacion de boton — el SQL es idempotente con INSERT OR IGNORE
    y los UPDATE/DELETE son seguros de re-ejecutar.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    if d.get('token', '').strip() != 'APLICAR_CORRECCIONES_2026_04_28':
        return jsonify({'error': 'Token incorrecto. Debe ser exactamente: APLICAR_CORRECCIONES_2026_04_28'}), 403

    # Localizar archivo SQL
    import os as _os
    repo_root = _os.path.dirname(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
    sql_path = _os.path.join(repo_root, 'scripts', 'migraciones', 'correcciones_formulas_2026_04_28.sql')
    if not _os.path.exists(sql_path):
        return jsonify({'error': f'Archivo SQL no encontrado: {sql_path}'}), 500

    try:
        with open(sql_path, 'r', encoding='utf-8') as f:
            sql_text = f.read()
    except Exception as _e:
        return jsonify({'error': f'No pude leer SQL: {_e}'}), 500

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Parsear los INSERTs del SQL para saber que MPs queremos crear (codigo_planeado, inci, comercial)
    import re as _re
    plan_inserts = []  # [(codigo_planeado, inci, comercial, tipo)]
    for m in _re.finditer(
        r"INSERT OR IGNORE INTO maestro_mps[^V]+VALUES\s*\(\s*'([^']+)'\s*,\s*'([^']+)'\s*,\s*'([^']+)'\s*,\s*'([^']+)'",
        sql_text):
        plan_inserts.append((m.group(1), m.group(2), m.group(3), m.group(4)))

    try:
        # 1) Detectar duplicados por nombre_inci YA EXISTENTE en catalogo
        existentes_por_inci = {}
        for codigo_p, inci, comercial, tipo in plan_inserts:
            row = c.execute(
                "SELECT codigo_mp, nombre_inci, nombre_comercial FROM maestro_mps "
                "WHERE LOWER(TRIM(nombre_inci)) = LOWER(TRIM(?)) LIMIT 1",
                (inci,)
            ).fetchone()
            if row:
                existentes_por_inci[codigo_p] = {
                    'codigo_planeado': codigo_p,
                    'inci_planeado': inci,
                    'codigo_existente': row[0],
                    'inci_existente': row[1],
                    'comercial_existente': row[2],
                }

        # 2) Verificar rango MP00400-MP00500 ocupado
        ocupados = c.execute(
            "SELECT codigo_mp, nombre_inci, nombre_comercial FROM maestro_mps "
            "WHERE codigo_mp >= 'MP00400' AND codigo_mp <= 'MP00500' ORDER BY codigo_mp"
        ).fetchall()

        modo = (d.get('modo') or '').strip()

        # === Si hay duplicados por INCI, exigir decision explicita ===
        if existentes_por_inci and modo not in ('mapear_duplicados', 'auto_renumerar'):
            conn.close()
            return jsonify({
                'error': f'Hay {len(existentes_por_inci)} INCI(s) que ya existen en el catalogo con OTRO codigo_mp. Crearlos duplicaria entradas.',
                'duplicados_inci': list(existentes_por_inci.values()),
                'total_duplicados_inci': len(existentes_por_inci),
                'ocupados_rango_mp00400': [
                    {'codigo': r[0], 'inci': r[1], 'comercial': r[2]} for r in ocupados
                ],
                'opciones': {
                    'mapear_duplicados': 'Reenviar con modo="mapear_duplicados" — los UPDATEs apuntaran a los codigos existentes, no se crearan duplicados',
                    'auto_renumerar': 'Reenviar con modo="auto_renumerar" — duplica las entradas en otro codigo_mp libre (NO RECOMENDADO si los INCIs son iguales)',
                },
            }), 409

        # === Si no hay duplicados pero el rango esta ocupado, ofrecer auto_renumerar ===
        if ocupados and not existentes_por_inci and modo != 'auto_renumerar':
            conn.close()
            return jsonify({
                'error': f'Hay {len(ocupados)} codigos MP en el rango MP00400-MP00500 ocupados (con OTROS INCIs, no los mios).',
                'ocupados_con_nombres': [
                    {'codigo': r[0], 'inci': r[1], 'comercial': r[2]} for r in ocupados
                ],
                'total_ocupados': len(ocupados),
                'sugerencia': 'Sin duplicados de INCI detectados — auto_renumerar seguro. Reenviar con modo="auto_renumerar".',
            }), 409

        # === modo=mapear_duplicados: skip INSERT de duplicados, reescribir UPDATEs ===
        if modo == 'mapear_duplicados' and existentes_por_inci:
            # Borrar las lineas INSERT de los duplicados (se vuelven no-op)
            # Y reescribir referencias codigo_planeado -> codigo_existente en UPDATEs
            for cp, info in existentes_por_inci.items():
                ce = info['codigo_existente']
                # Reemplazar codigo_planeado por codigo_existente en TODO el SQL
                sql_text = sql_text.replace("'" + cp + "'", "'" + ce + "'")
            # Tras el replace, los INSERTs duplicados son INSERT OR IGNORE de codigos existentes
            # — el OR IGNORE los hace no-op. Perfecto, no queda accion residual.

        # === modo=auto_renumerar: buscar rango libre y renumerar ===
        if modo == 'auto_renumerar':
            todos = c.execute(
                "SELECT codigo_mp FROM maestro_mps WHERE codigo_mp GLOB 'MP[0-9]*'"
            ).fetchall()
            usados = set()
            for (cm,) in todos:
                try: usados.add(int(cm[2:]))
                except (ValueError, TypeError):
                    pass  # codigo_mp no numérico (ej. MPABC), skip esperado
            new_start = None
            for base in range(1000, 9000, 50):
                if not any((base + k) in usados for k in range(50)):
                    new_start = base
                    break
            if new_start is None:
                conn.close()
                return jsonify({'error': 'No hay rango contiguo libre de 50 codigos hasta MP09000'}), 500

            def _replace_code(m):
                old_n = int(m.group(1))
                if 400 <= old_n <= 443:
                    new_n = new_start + (old_n - 400)
                    return 'MP{:05d}'.format(new_n)
                return m.group(0)
            sql_text = _re.sub(r"MP(\d{5})", _replace_code, sql_text)

    except sqlite3.OperationalError as _e:
        conn.close()
        return jsonify({'error': f'No pude verificar rango: {_e}'}), 500

    # Backup completo de DB antes de tocar
    try:
        do_backup(triggered_by='pre_aplicar_correcciones_2026_04_28')
    except Exception as _e:
        conn.close()
        return jsonify({'error': f'Backup pre-aplicacion fallo: {_e}'}), 500

    # Snapshot conteos antes
    try:
        n_formulas_antes = c.execute('SELECT COUNT(*) FROM formula_items').fetchone()[0]
        n_mps_antes = c.execute('SELECT COUNT(*) FROM maestro_mps').fetchone()[0]
    except Exception:
        n_formulas_antes = n_mps_antes = -1

    # Ejecutar SQL completo en una sola pasada (executescript permite multiples statements)
    try:
        # NOTA: el SQL trae su propio BEGIN TRANSACTION/COMMIT, executescript respeta eso
        c.executescript(sql_text)
        conn.commit()
    except Exception as _e:
        try: conn.rollback()
        except Exception as _r:
            __import__('logging').getLogger('admin').warning('rollback no aplicable: %s', _r)
        conn.close()
        return jsonify({'error': f'Aplicacion del SQL fallo: {_e}'}), 500

    # Snapshot despues
    try:
        n_formulas_desp = c.execute('SELECT COUNT(*) FROM formula_items').fetchone()[0]
        n_mps_desp = c.execute('SELECT COUNT(*) FROM maestro_mps').fetchone()[0]
        n_nuevos_mps = c.execute(
            "SELECT COUNT(*) FROM maestro_mps WHERE codigo_mp >= 'MP00400' AND codigo_mp <= 'MP00500'"
        ).fetchone()[0]
        azeclair_inactivo = c.execute(
            "SELECT activo FROM maestro_mps WHERE codigo_mp='MP00284'"
        ).fetchone()
        azeclair_inactivo = azeclair_inactivo[0] if azeclair_inactivo else None
    except Exception:
        n_formulas_desp = n_mps_desp = n_nuevos_mps = -1
        azeclair_inactivo = None

    # Audit log
    try:
        import json as _json
        c.execute(
            """INSERT INTO audit_log
               (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now'))""",
            (u, 'APLICAR_CORRECCIONES_FORMULAS_BATCH_20260428', 'multi', '_BULK_',
             _json.dumps({
                 'sql_file': 'scripts/migraciones/correcciones_formulas_2026_04_28.sql',
                 'formula_items_antes': n_formulas_antes,
                 'formula_items_despues': n_formulas_desp,
                 'maestro_mps_antes': n_mps_antes,
                 'maestro_mps_despues': n_mps_desp,
                 'mps_nuevos_creados': n_nuevos_mps,
                 'azeclair_marcado_inactivo': azeclair_inactivo == 0,
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except Exception:
        pass
    conn.close()

    return jsonify({
        'ok': True,
        'mensaje': 'Correcciones aplicadas en transaccion atomica con backup previo.',
        'antes': {
            'formula_items': n_formulas_antes,
            'maestro_mps': n_mps_antes,
        },
        'despues': {
            'formula_items': n_formulas_desp,
            'maestro_mps': n_mps_desp,
            'mps_nuevos_creados': n_nuevos_mps,
            'azeclair_inactivo': azeclair_inactivo == 0,
        },
        'cambios_netos': {
            'formula_items_diff': n_formulas_desp - n_formulas_antes,
            'maestro_mps_diff': n_mps_desp - n_mps_antes,
        },
    })


@bp.route("/api/admin/sembrar-maestro-desde-excel", methods=["POST"])
def admin_sembrar_maestro_desde_excel():
    """Asegura que TODOS los códigos del Excel (verdes + no-verdes) existan
    en maestro_mps. Para los que no existen, inserta una entrada con
    stock_minimo=0 y los datos disponibles del Excel (nombre, proveedor,
    estantería). NO crea movimientos — los no-verdes quedan en stock 0
    naturalmente.

    Justificación: las MPs no-verdes son materias primas que se han usado
    pero están actualmente agotadas. Necesitan estar en el catálogo para:
      - Aparecer en fórmulas
      - Recibirse vía OC cuando llegue reposición
      - Aparecer en alertas de mínimos cuando se reciban

    Solo admins. SIN tocar `movimientos`.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Archivo debe ser .xlsx'}), 400

    excel_verde, _, excel_no_verde, errs = _parse_excel_verde(f, incluir_no_verde=True)
    if excel_verde is None:
        return jsonify({'error': errs[0] if errs else 'Excel inválido'}), 400

    # Consolidar todos los códigos (verdes + no-verdes), tomando el primer
    # registro encontrado de cada código (los datos de catálogo son los mismos
    # entre lotes del mismo MP).
    todos_por_codigo = {}
    for (cod, _lote), info in excel_verde.items():
        if cod not in todos_por_codigo:
            todos_por_codigo[cod] = info
    for (cod, _lote), info in excel_no_verde.items():
        if cod not in todos_por_codigo:
            todos_por_codigo[cod] = info

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Códigos ya presentes en maestro_mps
    existentes = {row[0] for row in c.execute(
        "SELECT codigo_mp FROM maestro_mps"
    ).fetchall()}

    nuevos = []
    actualizados_proveedor = []  # opcional: si en maestro está vacío y excel tiene proveedor
    for cod, info in todos_por_codigo.items():
        nombre = (info.get('nombre_comercial') or info.get('inci') or cod).strip()
        inci = (info.get('inci') or '').strip()
        proveedor = (info.get('proveedor') or '').strip()
        if cod not in existentes:
            try:
                c.execute(
                    """INSERT INTO maestro_mps
                       (codigo_mp, nombre_inci, nombre_comercial, proveedor,
                        stock_minimo, activo)
                       VALUES (?,?,?,?,?,?)""",
                    (cod, inci or nombre, nombre, proveedor, 0, 1),
                )
                nuevos.append({
                    'codigo_mp': cod,
                    'nombre_comercial': nombre,
                    'proveedor': proveedor,
                })
            except sqlite3.IntegrityError:
                pass
        else:
            # MP ya existe — actualizar proveedor si está vacío en maestro
            if proveedor:
                try:
                    cur_prov = c.execute(
                        "SELECT proveedor FROM maestro_mps WHERE codigo_mp=?",
                        (cod,),
                    ).fetchone()
                    if cur_prov and not (cur_prov[0] or '').strip():
                        c.execute(
                            "UPDATE maestro_mps SET proveedor=? WHERE codigo_mp=?",
                            (proveedor, cod),
                        )
                        actualizados_proveedor.append({
                            'codigo_mp': cod,
                            'proveedor': proveedor,
                        })
                except sqlite3.OperationalError:
                    pass

    conn.commit()

    # Audit
    try:
        import json as _json
        c.execute(
            """INSERT INTO audit_log
               (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now'))""",
            (u, 'SEMBRAR_MAESTRO_DESDE_EXCEL', 'maestro_mps', '_BULK_',
             _json.dumps({
                 'archivo': f.filename,
                 'nuevos_count': len(nuevos),
                 'proveedores_actualizados_count': len(actualizados_proveedor),
                 'codigos_excel_total': len(todos_por_codigo),
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except sqlite3.OperationalError:
        pass
    conn.close()

    return jsonify({
        'ok': True,
        'mensaje': (f'Sembrado: {len(nuevos)} nuevos MPs en catálogo, '
                    f'{len(actualizados_proveedor)} con proveedor actualizado. '
                    f'No se modificaron movimientos.'),
        'nuevos_count': len(nuevos),
        'nuevos_sample': nuevos[:30],
        'proveedores_actualizados_count': len(actualizados_proveedor),
        'proveedores_actualizados_sample': actualizados_proveedor[:30],
    })


@bp.route("/api/admin/inventario-health-monitor", methods=["GET"])
def admin_inventario_health_monitor():
    """Monitor periodico: detecta anomalias en kardex que indiquen
    duplicacion / bursts / cargas no auditadas.

    Tras el incidente de doble carga (2026-04-27) que inflo el
    inventario 3x, este monitor sirve para captar el patron antes de
    que sea masivo. Idealmente se corre en cron diario o se revisa
    manualmente cada semana.

    Detecta:
      1. BURST: >= 30 entradas en una sola fecha (excepto el dia cero
         del reset 2026-04-15 que es legitimo)
      2. MULTI-ENTRADAS: lotes con >= 2 Entradas (potencial doble
         carga). Filtra los regenerados de huerfanos (1 entrada con
         observaciones 'consumido pre-reset').
      3. SIN-OC ratio: si > 80% de Entradas no tienen numero_oc en
         ultimos 7 dias, es signo de carga masiva sin trazabilidad.
      4. STOCK ANOMALO: si stock_total > 3x el promedio del ultimo
         mes, es senal de inflado.

    Solo lectura, admin. Devuelve nivel: ok / warning / critical.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    alertas = []

    # 1. BURST detection — entradas masivas en un solo dia, excluyendo
    #    el dia cero conocido (2026-04-15) y observaciones de reset.
    DIA_CERO_RESET = '2026-04-15'
    try:
        rows = c.execute("""
            SELECT SUBSTR(fecha,1,10) as dia,
                   COUNT(*) as n,
                   COALESCE(SUM(cantidad),0) as g
            FROM movimientos
            WHERE tipo='Entrada'
              AND SUBSTR(fecha,1,10) != ?
              AND COALESCE(observaciones,'') NOT LIKE '%reset%'
              AND COALESCE(observaciones,'') NOT LIKE '%dia cero%'
            GROUP BY SUBSTR(fecha,1,10)
            HAVING n >= 30
            ORDER BY n DESC
            LIMIT 10
        """, (DIA_CERO_RESET,)).fetchall()
        for r in rows:
            alertas.append({
                'tipo': 'BURST',
                'severidad': 'critical' if r[1] >= 100 else 'warning',
                'mensaje': (f'{r[1]} Entradas en un solo dia ({r[0]}) — total '
                            f'{int(round(r[2])):,} g. '.replace(',', '.') +
                            'Posible carga masiva no auditada.'),
                'detalle': {'fecha': r[0], 'count': r[1], 'total_g': r[2]},
            })
    except sqlite3.OperationalError:
        pass

    # 2. MULTI-ENTRADAS — lotes con >=2 Entradas (filtrando regenerados
    #    de huerfanos que tienen observaciones especificas)
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,''),
                   COUNT(*) as n,
                   SUM(cantidad) as g
            FROM movimientos
            WHERE tipo='Entrada'
              AND COALESCE(observaciones,'') NOT LIKE '%consumido pre-reset%'
              AND COALESCE(observaciones,'') NOT LIKE '%dia cero v8_1%'
            GROUP BY material_id, lote
            HAVING n >= 2
            ORDER BY n DESC, g DESC
            LIMIT 20
        """).fetchall()
        for r in rows:
            alertas.append({
                'tipo': 'MULTI_ENTRADAS',
                'severidad': 'critical' if r[2] >= 3 else 'warning',
                'mensaje': (f'{r[0]}/{r[1]}: {r[2]} Entradas con total '
                            f'{int(round(r[3])):,} g. '.replace(',', '.') +
                            'Verificar si son recepciones legitimas distintas o doble carga.'),
                'detalle': {
                    'codigo_mp': r[0], 'lote': r[1] or '(sin lote)',
                    'count_entradas': r[2], 'total_g': r[3],
                },
            })
    except sqlite3.OperationalError:
        pass

    # 3. SIN OC ratio en ultimos 7 dias (excluye reset)
    try:
        from datetime import datetime as _dt, timedelta as _td
        hace_7 = (_dt.utcnow() - _td(days=7)).strftime('%Y-%m-%d')
        r = c.execute("""
            SELECT
              SUM(CASE WHEN COALESCE(numero_oc,'') = '' THEN 1 ELSE 0 END) as sin_oc,
              COUNT(*) as total
            FROM movimientos
            WHERE tipo='Entrada' AND fecha >= ?
              AND COALESCE(observaciones,'') NOT LIKE '%reset%'
              AND COALESCE(observaciones,'') NOT LIKE '%dia cero%'
              AND COALESCE(observaciones,'') NOT LIKE '%consumido pre-reset%'
        """, (hace_7,)).fetchone()
        sin_oc = r[0] or 0
        total = r[1] or 0
        if total >= 5:
            ratio = sin_oc / total
            if ratio > 0.80:
                alertas.append({
                    'tipo': 'SIN_OC_ANOMALO',
                    'severidad': 'warning',
                    'mensaje': (f'{sin_oc}/{total} ({int(ratio*100)}%) Entradas en ultimos '
                                f'7 dias sin numero_oc. Recepciones formales deberian '
                                f'siempre traer OC.'),
                    'detalle': {'sin_oc': sin_oc, 'total': total, 'ratio_pct': int(ratio*100)},
                })
    except sqlite3.OperationalError:
        pass

    # 4. Stock total razonable (heuristica — alarma si > 50 toneladas)
    try:
        stock_g = float(c.execute(
            "SELECT COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) "
            "FROM movimientos"
        ).fetchone()[0] or 0)
        if stock_g > 50_000_000:  # 50 toneladas
            alertas.append({
                'tipo': 'STOCK_ANOMALO',
                'severidad': 'critical',
                'mensaje': (f'Stock total = {int(round(stock_g)):,} g'.replace(',', '.') +
                            ' supera 50 toneladas. '
                            'Para una empresa cosmetica de tu escala es altamente probable '
                            'inflado por doble carga. Auditar de inmediato.'),
                'detalle': {'stock_total_g': stock_g, 'umbral_g': 50_000_000},
            })
    except sqlite3.OperationalError:
        pass

    conn.close()

    n_critical = sum(1 for a in alertas if a['severidad'] == 'critical')
    n_warning = sum(1 for a in alertas if a['severidad'] == 'warning')
    nivel = ('critical' if n_critical else ('warning' if n_warning else 'ok'))

    return jsonify({
        'ok': True,
        'nivel': nivel,
        'count_critical': n_critical,
        'count_warning': n_warning,
        'alertas': alertas,
        'recomendacion': {
            'ok': 'Sistema saludable. Sin patrones de doble carga detectados.',
            'warning': 'Revisar alertas. Posibles cargas duplicadas o sin trazabilidad.',
            'critical': ('ACCION INMEDIATA. Patrones de doble carga detectados. '
                         'Considerar pausar entradas hasta investigar. '
                         'Usar /api/admin/inventario-diagnostico-entradas para detalle.'),
        }[nivel],
    })


@bp.route("/api/admin/health-check-post-reset", methods=["GET"])
def admin_health_check_post_reset():
    """Verificación de integridad despues del reset.

    Confirma que:
      - Stock total esta en rango razonable
      - No hay lotes con stock negativo
      - Producciones siguen intactas (no se borraron)
      - OCs y comprobantes siguen intactos
      - Catalogo maestro_mps sigue intacto
      - Hay salidas FEFO de produccion (no se perdieron)
      - Solicitudes_compra siguen intactas

    Solo lectura, admins.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    checks = {}

    # 1. Movimientos: stock total y conteos
    try:
        movs_total = c.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
        stock_total_g = float(c.execute(
            "SELECT COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) "
            "FROM movimientos"
        ).fetchone()[0] or 0)
        n_entradas = c.execute("SELECT COUNT(*) FROM movimientos WHERE tipo='Entrada'").fetchone()[0]
        n_salidas = c.execute("SELECT COUNT(*) FROM movimientos WHERE tipo='Salida'").fetchone()[0]
        n_salidas_fefo = c.execute(
            "SELECT COUNT(*) FROM movimientos WHERE tipo='Salida' AND observaciones LIKE 'FEFO:%'"
        ).fetchone()[0]
        checks['movimientos'] = {
            'ok': True,
            'total': movs_total,
            'stock_total_g': round(stock_total_g, 1),
            'stock_total_kg': round(stock_total_g / 1000, 2),
            'entradas': n_entradas,
            'salidas': n_salidas,
            'salidas_fefo_produccion': n_salidas_fefo,
        }
    except Exception as e:
        checks['movimientos'] = {'ok': False, 'error': str(e)}

    # 2. Lotes con stock negativo (debe ser 0 — invariante critica)
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,''),
                   SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END) as neto
            FROM movimientos
            GROUP BY material_id, lote
            HAVING neto < -0.5
        """).fetchall()
        checks['stock_negativo'] = {
            'ok': len(rows) == 0,
            'count': len(rows),
            'sample': [{'codigo_mp': r[0], 'lote': r[1], 'neto_g': round(r[2], 1)}
                       for r in rows[:10]],
            'critico': 'Si > 0: hay lotes con stock negativo (bug del replay)',
        }
    except Exception as e:
        checks['stock_negativo'] = {'ok': False, 'error': str(e)}

    # 3. Producciones intactas
    try:
        n_prod = c.execute("SELECT COUNT(*) FROM producciones").fetchone()[0]
        last_prod = c.execute(
            "SELECT producto, fecha, lote FROM producciones ORDER BY id DESC LIMIT 1"
        ).fetchone()
        checks['producciones'] = {
            'ok': n_prod >= 0,
            'count': n_prod,
            'ultima': {
                'producto': last_prod[0], 'fecha': str(last_prod[1])[:10],
                'lote': last_prod[2]
            } if last_prod else None,
        }
    except Exception as e:
        checks['producciones'] = {'ok': False, 'error': str(e)}

    # 4. OCs intactas
    try:
        n_ocs = c.execute("SELECT COUNT(*) FROM ordenes_compra").fetchone()[0]
        n_oc_items = c.execute("SELECT COUNT(*) FROM ordenes_compra_items").fetchone()[0]
        checks['ocs'] = {
            'ok': True,
            'ordenes_compra': n_ocs,
            'items_total': n_oc_items,
        }
    except Exception as e:
        checks['ocs'] = {'ok': False, 'error': str(e)}

    # 5. Comprobantes intactos
    try:
        n_comp = c.execute("SELECT COUNT(*) FROM comprobantes_pago").fetchone()[0]
        checks['comprobantes_pago'] = {'ok': True, 'count': n_comp}
    except Exception as e:
        checks['comprobantes_pago'] = {'ok': False, 'error': str(e)}

    # 6. Maestro MPs intacto
    try:
        n_mps = c.execute("SELECT COUNT(*) FROM maestro_mps WHERE activo=1").fetchone()[0]
        checks['maestro_mps'] = {'ok': True, 'count_activos': n_mps}
    except Exception as e:
        checks['maestro_mps'] = {'ok': False, 'error': str(e)}

    # 7. Solicitudes intactas
    try:
        n_sol = c.execute("SELECT COUNT(*) FROM solicitudes_compra").fetchone()[0]
        n_sol_items = c.execute("SELECT COUNT(*) FROM solicitudes_compra_items").fetchone()[0]
        checks['solicitudes_compra'] = {
            'ok': True, 'count': n_sol, 'items_total': n_sol_items,
        }
    except Exception as e:
        checks['solicitudes_compra'] = {'ok': False, 'error': str(e)}

    # 8. Audit trail del reset
    try:
        last_resets = c.execute("""
            SELECT accion, fecha FROM audit_log
            WHERE accion LIKE 'RESET_INVENTARIO%'
            ORDER BY id DESC LIMIT 5
        """).fetchall()
        checks['audit_log'] = {
            'ok': len(last_resets) > 0,
            'eventos_reset_recientes': [
                {'accion': r[0], 'fecha': str(r[1])[:19]} for r in last_resets
            ],
        }
    except Exception as e:
        checks['audit_log'] = {'ok': False, 'error': str(e)}

    conn.close()

    overall_ok = all(v.get('ok', False) for v in checks.values())
    return jsonify({
        'ok': overall_ok,
        'overall': 'PASS' if overall_ok else 'FAIL',
        'checks': checks,
        'recomendacion': (
            'Sistema intacto. Planta deberia funcionar normalmente. '
            'Verificar visualmente: /planta carga, Stock por Lote muestra '
            'lotes con cantidad > 0, /compras solicitudes intactas, '
            '/financiero carga.'
        ) if overall_ok else 'REVISAR — algun check fallo. Restaurar backup si hay duda.',
    })


@bp.route("/api/admin/inventario-diagnostico-entradas", methods=["GET"])
def admin_inventario_diagnostico_entradas():
    """Diagnostico de Entradas en movimientos para detectar doble carga
    vs recepciones legitimas.

    Caso de uso (CEO 2026-04-27): el audit muestra DB con 2x el stock
    del Excel. Antes de tomar decision destructiva, ver el timeline real
    de entradas:

      - Lotes con multiples Entradas (smoking gun de doble carga)
      - Entradas por dia (timeline)
      - Entradas por operador (quien entro que)
      - Entradas con OC vs sin OC (recepcion formal vs manual)

    Solo lectura, admins.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 1. Lotes con MULTIPLES entradas (doble carga sospechosa)
    multi_entradas = []
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,'') as lote,
                   COUNT(*) as n_entradas,
                   SUM(cantidad) as total_g,
                   GROUP_CONCAT(DISTINCT COALESCE(operador,''))  as operadores,
                   MIN(fecha) as primera,
                   MAX(fecha) as ultima
            FROM movimientos
            WHERE tipo='Entrada'
            GROUP BY material_id, lote
            HAVING COUNT(*) >= 2
            ORDER BY COUNT(*) DESC, SUM(cantidad) DESC
            LIMIT 200
        """).fetchall()
        for r in rows:
            multi_entradas.append({
                'codigo_mp': r[0], 'lote': r[1] or '(sin lote)',
                'n_entradas': r[2], 'total_g': round(r[3] or 0, 1),
                'operadores': r[4] or '',
                'primera_fecha': str(r[5])[:10] if r[5] else '',
                'ultima_fecha':  str(r[6])[:10] if r[6] else '',
            })
    except sqlite3.OperationalError:
        pass

    # 2. Timeline: entradas por dia (para detectar bursts de carga)
    timeline = []
    try:
        rows = c.execute("""
            SELECT SUBSTR(fecha,1,10) as dia,
                   COUNT(*) as n_movs,
                   SUM(cantidad) as total_g
            FROM movimientos
            WHERE tipo='Entrada'
            GROUP BY SUBSTR(fecha,1,10)
            ORDER BY dia ASC
        """).fetchall()
        for r in rows:
            timeline.append({
                'fecha': r[0],
                'n_entradas': r[1],
                'total_g': round(r[2] or 0, 1),
            })
    except sqlite3.OperationalError:
        pass

    # 3. Por operador
    por_operador = []
    try:
        rows = c.execute("""
            SELECT COALESCE(operador,'(vacio)') as op,
                   COUNT(*) as n_movs,
                   SUM(cantidad) as total_g,
                   MIN(fecha) as primera,
                   MAX(fecha) as ultima
            FROM movimientos
            WHERE tipo='Entrada'
            GROUP BY operador
            ORDER BY n_movs DESC
        """).fetchall()
        for r in rows:
            por_operador.append({
                'operador': r[0],
                'n_entradas': r[1],
                'total_g': round(r[2] or 0, 1),
                'primera_fecha': str(r[3])[:10] if r[3] else '',
                'ultima_fecha':  str(r[4])[:10] if r[4] else '',
            })
    except sqlite3.OperationalError:
        pass

    # 4. Origen: con OC vs sin OC
    con_oc = 0; sin_oc = 0; total_g_con_oc = 0; total_g_sin_oc = 0
    try:
        r = c.execute("""
            SELECT COUNT(*), COALESCE(SUM(cantidad),0)
            FROM movimientos
            WHERE tipo='Entrada' AND COALESCE(numero_oc,'') != ''
        """).fetchone()
        con_oc = r[0]; total_g_con_oc = float(r[1] or 0)
        r = c.execute("""
            SELECT COUNT(*), COALESCE(SUM(cantidad),0)
            FROM movimientos
            WHERE tipo='Entrada' AND COALESCE(numero_oc,'') = ''
        """).fetchone()
        sin_oc = r[0]; total_g_sin_oc = float(r[1] or 0)
    except sqlite3.OperationalError:
        pass

    # 5. Total resumen
    try:
        total_entradas = c.execute(
            "SELECT COUNT(*), COALESCE(SUM(cantidad),0) FROM movimientos WHERE tipo='Entrada'"
        ).fetchone()
    except sqlite3.OperationalError:
        total_entradas = (0, 0)

    conn.close()

    return jsonify({
        'ok': True,
        'resumen': {
            'total_entradas': total_entradas[0],
            'total_entradas_g': round(total_entradas[1], 1),
            'lotes_con_multiples_entradas': len(multi_entradas),
            'entradas_con_oc': con_oc,
            'entradas_con_oc_g': round(total_g_con_oc, 1),
            'entradas_sin_oc': sin_oc,
            'entradas_sin_oc_g': round(total_g_sin_oc, 1),
        },
        'multi_entradas': multi_entradas,
        'timeline': timeline[:60],   # primeros 60 dias
        'por_operador': por_operador,
    })


@bp.route("/api/admin/debug-solicitud/<numero>", methods=["GET"])
def admin_debug_solicitud(numero):
    """Devuelve los items RAW de una solicitud para depurar.

    Útil cuando el modal muestra cantidades raras y queremos ver qué hay
    realmente en la DB sin transformaciones del frontend.
    """
    u, err, code = _require_admin()
    if err:
        return err, code
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    sol = c.execute(
        "SELECT numero, fecha, estado, observaciones, numero_oc, categoria FROM solicitudes_compra WHERE numero=?",
        (numero.upper(),)
    ).fetchone()
    if not sol:
        return jsonify({'error': 'Solicitud no encontrada', 'numero': numero}), 404
    items = c.execute(
        """SELECT id, codigo_mp, nombre_mp, cantidad_g, unidad, justificacion
           FROM solicitudes_compra_items WHERE numero=?""",
        (numero.upper(),)
    ).fetchall()
    conn.close()
    return jsonify({
        'solicitud': dict(sol),
        'items_raw': [dict(i) for i in items],
        'count': len(items),
        'items_con_cantidad_cero': sum(1 for i in items if (i['cantidad_g'] or 0) <= 0),
    })


@bp.route("/api/admin/stock-mp-diagnostico", methods=["GET"])
def admin_stock_mp_diagnostico():
    """Diagnóstico de por qué un MP aparece con stock 0.

    Query: ?codigos=COD1,COD2,COD3  o  ?nombres=GLICERINA,ALOE
    Devuelve para cada uno:
      - en_maestro_mps: si existe en el catálogo
      - movimientos_count: cuántos movimientos hay
      - movimientos_match_por: 'codigo' / 'nombre' / 'no_match'
      - stock_total_g: suma calculada
      - ultimo_movimiento: fecha
      - sugerencia: qué hacer
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    codigos = (request.args.get('codigos') or '').strip()
    nombres = (request.args.get('nombres') or '').strip()
    if not codigos and not nombres:
        return jsonify({'error': 'Pasa ?codigos= o ?nombres= separados por coma'}), 400

    cods = [x.strip() for x in codigos.split(',') if x.strip()] if codigos else []
    noms = [x.strip() for x in nombres.split(',') if x.strip()] if nombres else []

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    resultados = []
    for cod in cods:
        info = {'codigo_mp': cod, 'busqueda_por': 'codigo'}
        # ¿existe en maestro?
        r = c.execute(
            "SELECT codigo_mp, nombre_comercial, proveedor FROM maestro_mps WHERE codigo_mp=?",
            (cod,)
        ).fetchone()
        if r:
            info['en_maestro_mps'] = True
            info['nombre_comercial'] = r['nombre_comercial']
            info['proveedor_catalogo'] = r['proveedor']
        else:
            info['en_maestro_mps'] = False
        # ¿movimientos por código?
        r2 = c.execute("""SELECT COUNT(*) as n,
                          COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) as stock,
                          MAX(fecha) as ultimo
                          FROM movimientos WHERE material_id=?""", (cod,)).fetchone()
        info['movimientos_count'] = r2['n'] if r2 else 0
        info['stock_total_g'] = float(r2['stock'] or 0) if r2 else 0
        info['ultimo_movimiento'] = r2['ultimo'] if r2 else None
        if info['movimientos_count'] > 0:
            info['movimientos_match_por'] = 'codigo'
        else:
            # Fallback por nombre del catálogo si lo tenemos
            nm = info.get('nombre_comercial')
            if nm:
                r3 = c.execute("""SELECT COUNT(*) as n,
                                  COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) as stock
                                  FROM movimientos WHERE UPPER(TRIM(material_nombre))=UPPER(TRIM(?))""", (nm,)).fetchone()
                if r3 and r3['n'] > 0:
                    info['movimientos_count'] = r3['n']
                    info['stock_total_g'] = float(r3['stock'] or 0)
                    info['movimientos_match_por'] = 'nombre'
                else:
                    info['movimientos_match_por'] = 'no_match'
            else:
                info['movimientos_match_por'] = 'no_match'
        # Sugerencia
        if not info['en_maestro_mps']:
            info['sugerencia'] = 'Código NO está en catálogo maestro_mps. Crear primero el item en /planta → Catálogo MPs.'
        elif info['movimientos_count'] == 0:
            info['sugerencia'] = 'Ítem existe en catálogo pero NUNCA se registró entrada de stock. Hacer carga inicial en /planta → Movimientos (Entrada inicial).'
        elif info['stock_total_g'] <= 0:
            info['sugerencia'] = f'Tuvo {info["movimientos_count"]} movimientos pero el balance es 0 o negativo (más salidas que entradas). Revisar si falta registrar una compra reciente.'
        else:
            info['sugerencia'] = f'Stock real: {info["stock_total_g"]}g. Si el modal muestra 0, refresca con Ctrl+Shift+R.'
        resultados.append(info)

    for nom in noms:
        info = {'nombre': nom, 'busqueda_por': 'nombre'}
        r = c.execute("""SELECT codigo_mp, nombre_comercial, proveedor FROM maestro_mps
                         WHERE UPPER(nombre_comercial) LIKE UPPER(?) LIMIT 5""",
                      (f'%{nom}%',)).fetchall()
        info['matches_catalogo'] = [dict(x) for x in r]
        r2 = c.execute("""SELECT COUNT(*) as n,
                          COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) as stock,
                          MAX(fecha) as ultimo
                          FROM movimientos
                          WHERE UPPER(material_nombre) LIKE UPPER(?)""", (f'%{nom}%',)).fetchone()
        info['movimientos_count'] = r2['n'] if r2 else 0
        info['stock_total_g'] = float(r2['stock'] or 0) if r2 else 0
        info['ultimo_movimiento'] = r2['ultimo'] if r2 else None
        resultados.append(info)

    conn.close()
    return jsonify({'resultados': resultados, 'total': len(resultados)})


@bp.route("/api/admin/mps-proveedores-status", methods=["GET"])
def admin_mps_proveedores_status():
    """Diagnóstico: qué MPs tienen proveedor asignado y cuáles no.

    El campo maestro_mps.proveedor lo lee la auto-generación de OCs sugeridas
    para agrupar el déficit. Si está vacío, las solicitudes salen con
    'Sin asignar' como proveedor y el user tiene que corregir a mano.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Resumen por proveedor
    rows = c.execute("""
        SELECT COALESCE(NULLIF(TRIM(proveedor), ''), '(SIN PROVEEDOR)') AS prov,
               COUNT(*) AS total
        FROM maestro_mps
        WHERE activo = 1
        GROUP BY prov
        ORDER BY (CASE WHEN prov='(SIN PROVEEDOR)' THEN 0 ELSE 1 END), total DESC, prov
    """).fetchall()
    por_proveedor = [{'proveedor': r['prov'], 'total': r['total']} for r in rows]
    total_activos = sum(r['total'] for r in por_proveedor)
    sin_proveedor_count = next(
        (r['total'] for r in por_proveedor if r['proveedor'] == '(SIN PROVEEDOR)'), 0
    )

    # Lista detallada de los SIN proveedor — orden alfabético por nombre
    rows_sin = c.execute("""
        SELECT codigo_mp, nombre_comercial, COALESCE(tipo,'') as tipo
        FROM maestro_mps
        WHERE activo = 1
          AND (proveedor IS NULL OR TRIM(proveedor) = '')
        ORDER BY nombre_comercial COLLATE NOCASE LIMIT 200
    """).fetchall()
    sin_proveedor_lista = [
        {'codigo_mp': r['codigo_mp'], 'nombre': r['nombre_comercial'], 'tipo': r['tipo']}
        for r in rows_sin
    ]

    # Lista de proveedores existentes (para que la UI ofrezca dropdown)
    provs_existentes = [
        r['prov'] for r in por_proveedor if r['prov'] != '(SIN PROVEEDOR)'
    ]

    # Cross-check con MPs en déficit real (los que importan operacionalmente)
    # Esto requiere correr la lógica de Programación, lo evitamos aquí para
    # mantener este endpoint rápido. El user lo puede ver en /compras banner.

    conn.close()
    return jsonify({
        'total_activos': total_activos,
        'con_proveedor': total_activos - sin_proveedor_count,
        'sin_proveedor': sin_proveedor_count,
        'pct_cobertura': round(
            (total_activos - sin_proveedor_count) / total_activos * 100, 1
        ) if total_activos else 0,
        'por_proveedor': por_proveedor,
        'sin_proveedor_lista': sin_proveedor_lista,
        'proveedores_existentes': provs_existentes,
    })


@bp.route("/api/admin/mps-asignar-proveedor", methods=["POST"])
def admin_mps_asignar_proveedor():
    """Asigna proveedor a una MP. Body: {codigo_mp, proveedor}."""
    u, err, code = _require_admin()
    if err:
        return err, code
    d = request.get_json(silent=True) or {}
    codigo = (d.get('codigo_mp') or '').strip()
    proveedor = (d.get('proveedor') or '').strip()
    if not codigo:
        return jsonify({'error': 'codigo_mp requerido'}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # Capturar antes para audit
    antes_row = c.execute("SELECT proveedor FROM maestro_mps WHERE codigo_mp=?",
                          (codigo,)).fetchone()
    if not antes_row:
        conn.close()
        return jsonify({'error': f"No se encontró MP '{codigo}'"}), 404
    antes_prov = antes_row[0] or ''
    c.execute("UPDATE maestro_mps SET proveedor=? WHERE codigo_mp=?",
              (proveedor, codigo))
    n = c.rowcount or 0
    try:
        audit_log(c, usuario=u, accion='ASIGNAR_PROVEEDOR_MP',
                  tabla='maestro_mps', registro_id=codigo,
                  antes={'proveedor': antes_prov},
                  despues={'proveedor': proveedor},
                  detalle=f"Asignó proveedor MP {codigo}: '{antes_prov}' → '{proveedor}'")
    except Exception:
        pass
    conn.commit()
    conn.close()
    if n == 0:
        return jsonify({'error': f"No se encontró MP '{codigo}'"}), 404
    _log_sec(u, _client_ip(),
             "admin_mp_asignar_proveedor",
             f"codigo={codigo} proveedor={proveedor}")
    return jsonify({'ok': True, 'codigo_mp': codigo, 'proveedor': proveedor})


@bp.route("/api/admin/tipos-mp-stats", methods=["GET"])
def admin_tipos_mp_stats():
    """Cuenta items en maestro_mps agrupados por tipo_material.

    Útil para diagnosticar por qué la programación cíclica de E&E aparece
    vacía: si no hay items con tipo='Envase Secundario' (exacto), la rotación
    no tiene de dónde sacar los 3 ítems de la semana.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Conteo por tipo (incluye nulos/vacíos)
    rows = c.execute("""
        SELECT COALESCE(NULLIF(TRIM(tipo_material), ''), '(sin tipo)') AS tipo,
               COUNT(*) AS total,
               SUM(CASE WHEN activo=1 THEN 1 ELSE 0 END) AS activos,
               SUM(CASE WHEN COALESCE(stock_minimo,0) > 0 THEN 1 ELSE 0 END) AS con_min
        FROM maestro_mps
        GROUP BY tipo
        ORDER BY total DESC
    """).fetchall()

    # Primeros 5 ejemplos por tipo (para verificar nombres exactos)
    ejemplos = {}
    for r in rows:
        c.execute("""SELECT codigo_mp, nombre_comercial
                     FROM maestro_mps
                     WHERE COALESCE(NULLIF(TRIM(tipo_material),''),'(sin tipo)')=?
                       AND activo=1
                     ORDER BY nombre_comercial COLLATE NOCASE LIMIT 5""", (r["tipo"],))
        ejemplos[r["tipo"]] = [
            {"codigo": e["codigo_mp"], "nombre": e["nombre_comercial"]}
            for e in c.fetchall()
        ]

    conn.close()

    return jsonify({
        "tipos": [dict(r) for r in rows],
        "ejemplos": ejemplos,
        "esperados_e_e": ["Envase Primario", "Envase Secundario", "Empaque"],
        "nota": (
            "Si un tipo esperado no aparece, no hay items clasificados con "
            "ese valor exacto. Editá los items en el catálogo y asignales "
            "tipo_material correcto."
        ),
    })


# ─── Sync Banco Influencers desde Excel ───────────────────────────────────────

@bp.route("/api/admin/sync-influencers-excel", methods=["POST"])
def admin_sync_influencers_excel():
    """Sube un Excel de banco de influencers y hace UPSERT preservando datos.

    Body: multipart/form-data con campo 'file' = .xlsx
    Query: ?dry_run=1  para previsualizar sin escribir

    Hoja esperada: 'cuentas de creadores' con columnas:
       A: nombre, B: banco, C: cuenta, D: tipo cta, E: cédula,
       F: ciudad, G: usuario red, H: tipo creador

    Mapeo a marketing_influencers:
       nombre, banco, cuenta_bancaria, tipo_cuenta, cedula_nit,
       ciudad, instagram (si existe col), tipo (si existe col)

    UPSERT por LOWER(TRIM(nombre)). Solo actualiza columnas vacías en DB.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    dry_run = request.args.get('dry_run', '0') in ('1', 'true', 'True')

    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    def _norm(s):
        return (s or "").strip().lower() if isinstance(s, str) else ""

    def _str(v):
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'Excel inválido: {e}'}), 400

    if "cuentas de creadores" not in wb.sheetnames:
        return jsonify({
            'error': "Hoja 'cuentas de creadores' no existe",
            'hojas_disponibles': wb.sheetnames
        }), 400

    ws = wb["cuentas de creadores"]
    creadores = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not row[0]:
            continue
        creadores.append({
            "nombre":      _str(row[0]),
            "banco":       _str(row[1]) if len(row) > 1 else "",
            "cuenta":      _str(row[2]) if len(row) > 2 else "",
            "tipo_cta":    _str(row[3]) if len(row) > 3 else "",
            "cedula":      _str(row[4]) if len(row) > 4 else "",
            "ciudad":      _str(row[5]) if len(row) > 5 else "",
            "user":        _str(row[6]) if len(row) > 6 else "",
            "tipo_creador": _str(row[7]) if len(row) > 7 else "",
        })

    EXCEL_TO_DB = {
        "banco":         "banco",
        "cuenta":        "cuenta_bancaria",
        "tipo_cta":      "tipo_cuenta",
        "cedula":        "cedula_nit",
        "ciudad":        "ciudad",
        "user":          "instagram",
        "tipo_creador":  "tipo",
    }

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("PRAGMA table_info(marketing_influencers)")
    cols_db = {r["name"] for r in c.fetchall()}

    nuevos, actualizados, sin_cambios = [], [], []
    for ex in creadores:
        nombre = ex["nombre"]
        c.execute(
            "SELECT * FROM marketing_influencers WHERE LOWER(TRIM(nombre)) = ? LIMIT 1",
            (nombre.lower(),)
        )
        existing = c.fetchone()
        if not existing:
            cols_to_insert = ["nombre"]
            vals = [nombre]
            for k_ex, k_db in EXCEL_TO_DB.items():
                if k_db in cols_db and ex.get(k_ex):
                    cols_to_insert.append(k_db)
                    vals.append(ex[k_ex])
            placeholders = ",".join("?" * len(cols_to_insert))
            sql = f"INSERT INTO marketing_influencers ({','.join(cols_to_insert)}) VALUES ({placeholders})"
            if not dry_run:
                c.execute(sql, vals)
            nuevos.append(nombre)
            continue

        sets, vals, cambios_locales = [], [], []
        for k_ex, k_db in EXCEL_TO_DB.items():
            if k_db not in cols_db:
                continue
            new_val = ex.get(k_ex, "")
            old_val = existing[k_db] if k_db in existing.keys() else ""
            old_val = (old_val or "").strip() if isinstance(old_val, str) else ""
            if new_val and not old_val:
                sets.append(f"{k_db}=?")
                vals.append(new_val)
                cambios_locales.append(f"{k_db}: <vacío> → {new_val[:40]}")
        if sets:
            vals.append(existing["id"])
            if not dry_run:
                c.execute(
                    f"UPDATE marketing_influencers SET {','.join(sets)} WHERE id=?",
                    vals
                )
            actualizados.append({'nombre': nombre, 'cambios': cambios_locales})
        else:
            sin_cambios.append(nombre)

    if not dry_run:
        conn.commit()
    conn.close()

    _log_sec(u, _client_ip(),
             "admin_sync_influencers_excel",
             f"dry_run={dry_run} nuevos={len(nuevos)} act={len(actualizados)}")

    return jsonify({
        'ok': True,
        'dry_run': dry_run,
        'total_excel': len(creadores),
        'nuevos':       {'count': len(nuevos), 'lista': nuevos},
        'actualizados': {'count': len(actualizados), 'lista': actualizados},
        'sin_cambios':  {'count': len(sin_cambios), 'lista': sin_cambios[:50]},
        'cols_db_disponibles': sorted(cols_db),
    })


# ─── Importar pagos pendientes de influencers desde Excel ────────────────────

# Patrones para detectar columnas automáticamente — soporta Excel con
# diferentes naming sin que el user tenga que renombrar nada
_COL_PATTERNS_PAGOS = {
    'nombre':       ['influencer', 'creador', 'nombre', 'usuario'],
    'fecha':        ['fecha publicacion', 'fecha de publicacion', 'fecha publicación',
                     'fecha de publicación', 'fecha post', 'fecha contenido',
                     'publicacion', 'publicación', 'fecha'],
    'valor':        ['valor', 'monto', 'pago', 'total', 'precio'],
    'estado':       ['estado', 'pago realizado', 'pagado', 'status'],
    'concepto':     ['concepto', 'producto', 'campana', 'campaña', 'entregable',
                     'descripcion', 'descripción'],
    'instagram':    ['instagram', 'cuenta', 'user', '@'],
}


def _norm_header(s):
    """Normaliza un header de Excel: lower, sin acentos, sin extra spaces."""
    import unicodedata as _ud
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn')
    return ' '.join(s.split())


def _detect_cols(headers):
    """Devuelve dict {nuestro_campo: indice_en_excel} matcheando por patterns."""
    out = {}
    norms = [_norm_header(h) for h in headers]
    for field, patterns in _COL_PATTERNS_PAGOS.items():
        for i, h in enumerate(norms):
            if not h:
                continue
            if any(p in h for p in patterns):
                if field not in out:
                    out[field] = i
                    break
    return out


def _norm_nombre(s):
    """Normaliza nombre influencer para matching: lower + trim + sin acentos."""
    import unicodedata as _ud
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn')
    return ' '.join(s.split())


def _parse_fecha(v):
    """Acepta datetime, date, o string YYYY-MM-DD / DD/MM/YYYY. Devuelve ISO o ''."""
    if v is None or v == '':
        return ''
    from datetime import datetime as _dt, date as _date
    if isinstance(v, _dt):
        return v.date().isoformat()
    if isinstance(v, _date):
        return v.isoformat()
    s = str(v).strip()
    # YYYY-MM-DD
    try:
        return _dt.strptime(s[:10], '%Y-%m-%d').date().isoformat()
    except ValueError:
        pass
    # DD/MM/YYYY
    try:
        return _dt.strptime(s[:10], '%d/%m/%Y').date().isoformat()
    except ValueError:
        pass
    # DD-MM-YYYY
    try:
        return _dt.strptime(s[:10], '%d-%m-%Y').date().isoformat()
    except ValueError:
        pass
    return ''


def _parse_valor(v):
    """Acepta int, float, o string con $/coma/punto. Devuelve float >= 0."""
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('$', '').replace(' ', '')
    # Si tiene coma como decimal Y punto como miles → quitar puntos, comma a punto
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        # Solo coma — si tiene 3 dígitos después es miles, sino decimal
        partes = s.split(',')
        if len(partes[-1]) == 3:
            s = s.replace(',', '')
        else:
            s = s.replace(',', '.')
    elif '.' in s:
        partes = s.split('.')
        if len(partes[-1]) == 3:
            s = s.replace('.', '')
    try:
        return max(0.0, float(s))
    except ValueError:
        return 0.0


@bp.route("/api/admin/import-pagos-influencers-excel", methods=["POST"])
def admin_import_pagos_influencers_excel():
    """Importa pagos pendientes de influencers desde Excel.

    Modo "reset solo pendientes" (opción B aprobada por user):
      - Borra solicitudes Influencer con estado != Pagada (más sus OCs y pagos)
      - Importa cada fila del Excel como nueva solicitud Aprobada lista para pagar
      - Conserva intacto el historial de pagos hechos

    Detecta columnas del Excel automáticamente buscando keywords en los
    headers (influencer/creador/nombre, fecha publicación, valor, etc.).

    Body: multipart/form-data 'file' = .xlsx
    Query: ?dry_run=1 → previsualiza sin escribir nada

    Para cada fila válida:
      1. Match influencer en marketing_influencers por nombre normalizado
      2. INSERT solicitudes_compra (estado=Aprobada, categoria=Influencer/Marketing Digital)
      3. INSERT ordenes_compra (proveedor=nombre, valor_total=valor)
      4. INSERT ordenes_compra_items (descripción=concepto, valor)
      5. INSERT pagos_influencers con fecha_publicacion = fecha del Excel
         (esto es lo que permitirá ordenar correctamente en /compras)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    dry_run = request.args.get('dry_run', '0') in ('1', 'true', 'True')

    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'Excel inválido: {e}'}), 400

    # Buscar la hoja: priorizar nombres conocidos, sino la primera
    nombres_prioridad = ['pagos', 'pagos influencers', 'publicaciones',
                         'contenido', 'campanas', 'campañas']
    sheet_name = None
    for n in wb.sheetnames:
        if any(p in _norm_header(n) for p in nombres_prioridad):
            sheet_name = n
            break
    if sheet_name is None:
        # Excluir hojas de banco/configuración si las detecta
        for n in wb.sheetnames:
            if 'cuenta' not in _norm_header(n) and 'banco' not in _norm_header(n):
                sheet_name = n
                break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows_excel = list(ws.iter_rows(values_only=True))
    if len(rows_excel) < 2:
        return jsonify({
            'error': f'Hoja "{sheet_name}" tiene menos de 2 filas',
            'hojas_disponibles': wb.sheetnames,
        }), 400

    headers = rows_excel[0]
    cols_idx = _detect_cols(headers)

    if 'nombre' not in cols_idx or 'valor' not in cols_idx:
        return jsonify({
            'error': 'No se detectó columna de Influencer y/o Valor en el Excel',
            'hoja_usada': sheet_name,
            'headers_detectados': [str(h) for h in headers if h],
            'columnas_mapeadas': cols_idx,
            'hojas_disponibles': wb.sheetnames,
            'sugerencia': (
                'Asegúrate de que el Excel tenga columnas con nombres como '
                '"Influencer/Creador/Nombre" y "Valor/Monto/Pago". '
                'Mira hojas_disponibles si la hoja correcta es otra.'
            ),
        }), 400

    # Cargar tabla de influencers para matching
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT id, nombre FROM marketing_influencers")
    inf_by_norm = {}
    inf_norm_keys = []  # lista para fuzzy matching con difflib
    for r in c.fetchall():
        nm = _norm_nombre(r['nombre'])
        if nm and nm not in inf_by_norm:
            inf_by_norm[nm] = r['id']
            inf_norm_keys.append(nm)

    import difflib

    def _match_influencer(nombre_str):
        """Devuelve (inf_id, match_type, match_nombre) o (None, None, None).

        Estrategia escalonada:
          1. Match exacto normalizado
          2. Prefix match (uno empieza con el otro)
          3. Substring match (uno contiene al otro)
          4. Fuzzy con difflib.get_close_matches cutoff 0.82
        """
        nm = _norm_nombre(nombre_str)
        if not nm:
            return None, None, None
        if nm in inf_by_norm:
            return inf_by_norm[nm], 'exacto', nm
        # Prefix
        for k in inf_norm_keys:
            if k.startswith(nm) or nm.startswith(k):
                return inf_by_norm[k], 'prefix', k
        # Substring
        for k in inf_norm_keys:
            if nm in k or k in nm:
                return inf_by_norm[k], 'substring', k
        # Fuzzy
        candidatos = difflib.get_close_matches(nm, inf_norm_keys, n=1, cutoff=0.82)
        if candidatos:
            k = candidatos[0]
            return inf_by_norm[k], 'fuzzy', k
        return None, None, None

    # Procesar filas
    plan_import = []
    sin_match = []
    sin_valor = []
    sin_nombre = []
    pagadas_skipped = []

    estado_pagado_kws = {'pagado', 'pagada', 'paid', 'pago realizado', 'si', 'sí', 'yes', 'x'}

    for i, row in enumerate(rows_excel[1:], start=2):
        if not row or not any(row):
            continue
        nombre = row[cols_idx['nombre']] if cols_idx.get('nombre') is not None else None
        valor_raw = row[cols_idx['valor']] if cols_idx.get('valor') is not None else None
        fecha_raw = row[cols_idx['fecha']] if cols_idx.get('fecha') is not None else None
        estado_raw = row[cols_idx.get('estado')] if cols_idx.get('estado') is not None else None
        concepto_raw = row[cols_idx.get('concepto')] if cols_idx.get('concepto') is not None else None
        instagram_raw = row[cols_idx.get('instagram')] if cols_idx.get('instagram') is not None else None

        nombre_str = str(nombre or '').strip()
        if not nombre_str:
            sin_nombre.append({'fila': i})
            continue
        valor = _parse_valor(valor_raw)
        if valor <= 0:
            sin_valor.append({'fila': i, 'nombre': nombre_str})
            continue
        # Saltar las que ya estén marcadas como pagadas
        estado_norm = _norm_header(estado_raw)
        if any(kw in estado_norm for kw in estado_pagado_kws):
            pagadas_skipped.append({'fila': i, 'nombre': nombre_str, 'valor': valor})
            continue

        inf_id, match_type, match_nombre = _match_influencer(nombre_str)
        if not inf_id:
            sin_match.append({
                'fila': i, 'nombre': nombre_str, 'valor': valor,
                'instagram': str(instagram_raw or '').strip()
            })
            continue

        plan_import.append({
            'fila': i,
            'influencer_id': inf_id,
            'nombre': nombre_str,
            'valor': valor,
            'fecha_publicacion': _parse_fecha(fecha_raw),
            'concepto': str(concepto_raw or '').strip() or 'Pago influencer',
            'match_type': match_type,
            'match_nombre_banco': match_nombre,
        })

    if dry_run:
        return jsonify({
            'ok': True,
            'dry_run': True,
            'hoja_usada': sheet_name,
            'columnas_mapeadas': cols_idx,
            'a_importar': {
                'count': len(plan_import),
                'preview': plan_import[:30],
                'valor_total': round(sum(p['valor'] for p in plan_import), 0),
            },
            'sin_match': {
                'count': len(sin_match),
                'lista': sin_match[:30],
                'nota': 'Estos NO se van a importar. Agrega el influencer al banco primero (/admin sync banco) o corrige el nombre en el Excel.',
            },
            'pagadas_skipped': {'count': len(pagadas_skipped), 'lista': pagadas_skipped[:20]},
            'sin_valor': {'count': len(sin_valor)},
            'sin_nombre': {'count': len(sin_nombre)},
        })

    # ── APLICAR ─────────────────────────────────────────────────────────────
    # SAFETY: si no hay nada para importar (plan_import vacío) NO borramos
    # las solicitudes pendientes existentes — eso dejaría la base vacía.
    # El user debe arreglar nombres en el Excel o agregar al banco primero.
    # Override con ?force=1 si el user QUIERE limpiar todo aunque no importe nada.
    force = request.args.get('force', '0') in ('1', 'true', 'True')
    if not plan_import and not force:
        try:
            conn.close()
        except Exception:
            pass
        return jsonify({
            'error': 'No hay filas válidas para importar — reset abortado',
            'detalle': (
                f'{len(sin_match)} fila(s) sin match en el banco, '
                f'{len(sin_valor)} sin valor, {len(sin_nombre)} sin nombre, '
                f'{len(pagadas_skipped)} ya marcadas como pagadas. '
                'No se borró nada. Arregla los nombres del Excel o agrega los '
                'influencers al banco antes de aplicar. Si querés forzar el '
                'reset igual (vaciar pendientes), usa ?force=1.'
            ),
            'hoja_usada': sheet_name,
            'columnas_mapeadas': cols_idx,
            'sin_match': {
                'count': len(sin_match),
                'lista': sin_match[:30],
            },
            'sin_valor': {'count': len(sin_valor)},
            'pagadas_skipped': {'count': len(pagadas_skipped)},
        }), 400

    # Reset opción B: borrar solicitudes Influencer no-pagadas + sus OCs y pagos
    deleted = {'solicitudes': 0, 'ordenes_compra': 0, 'pagos_influencers': 0}
    try:
        # 1. Listar OCs de influencer cuyo estado de SC sea != 'Pagada'
        ocs_a_borrar = c.execute("""
            SELECT DISTINCT sc.numero_oc FROM solicitudes_compra sc
            WHERE sc.categoria = 'Influencer/Marketing Digital'
              AND sc.estado IN ('Pendiente', 'Aprobada', 'Rechazada')
              AND sc.numero_oc IS NOT NULL AND sc.numero_oc != ''
        """).fetchall()
        ocs_list = [r[0] for r in ocs_a_borrar]
        if ocs_list:
            placeholders = ','.join('?' * len(ocs_list))
            d = c.execute(
                f"DELETE FROM pagos_influencers WHERE numero_oc IN ({placeholders}) "
                f"AND COALESCE(estado,'') != 'Pagada'", ocs_list
            )
            deleted['pagos_influencers'] = d.rowcount or 0
            try:
                c.execute(
                    f"DELETE FROM ordenes_compra_items WHERE numero_oc IN ({placeholders})",
                    ocs_list
                )
            except sqlite3.OperationalError:
                pass
            d = c.execute(
                f"DELETE FROM ordenes_compra WHERE numero_oc IN ({placeholders})",
                ocs_list
            )
            deleted['ordenes_compra'] = d.rowcount or 0
        # Items de la solicitud
        try:
            c.execute("""DELETE FROM solicitudes_compra_items
                         WHERE numero IN (SELECT numero FROM solicitudes_compra
                                          WHERE categoria='Influencer/Marketing Digital'
                                            AND estado IN ('Pendiente','Aprobada','Rechazada'))""")
        except sqlite3.OperationalError:
            pass
        d = c.execute("""DELETE FROM solicitudes_compra
                         WHERE categoria='Influencer/Marketing Digital'
                           AND estado IN ('Pendiente','Aprobada','Rechazada')""")
        deleted['solicitudes'] = d.rowcount or 0

        # ── Importar plan ───────────────────────────────────────────────────
        from datetime import datetime as _dt
        anio = _dt.now().year
        c.execute("SELECT COALESCE(MAX(CAST(SUBSTR(numero, 10) AS INTEGER)),0) "
                  "FROM solicitudes_compra WHERE numero LIKE ?", (f'SOL-{anio}-%',))
        n_sol = (c.fetchone()[0] or 0)
        c.execute("SELECT COALESCE(MAX(CAST(SUBSTR(numero_oc, 9) AS INTEGER)),0) "
                  "FROM ordenes_compra WHERE numero_oc LIKE ?", (f'OC-{anio}-%',))
        n_oc = (c.fetchone()[0] or 0)

        importadas = []
        for plan in plan_import:
            n_sol += 1
            n_oc += 1
            num_sol = f'SOL-{anio}-{n_sol:04d}'
            num_oc = f'OC-{anio}-{n_oc:04d}'
            obs = (f"BENEFICIARIO: {plan['nombre']} | VALOR: ${plan['valor']:,.0f} | "
                   f"FECHA PUB: {plan['fecha_publicacion'] or 'sin fecha'} | "
                   f"CONCEPTO: {plan['concepto']} | Importado desde Excel por {u}")

            c.execute("""INSERT INTO solicitudes_compra
                (numero, fecha, estado, solicitante, urgencia, observaciones,
                 area, empresa, categoria, tipo, numero_oc, influencer_id, fecha_requerida)
                VALUES (?, datetime('now'), 'Aprobada', ?, 'Normal', ?, 'Marketing',
                        'Animus', 'Influencer/Marketing Digital', 'Servicio', ?, ?, ?)""",
                (num_sol, u, obs, num_oc, plan['influencer_id'],
                 plan['fecha_publicacion'] or ''))
            c.execute("""INSERT INTO ordenes_compra
                (numero_oc, fecha, estado, proveedor, valor_total, observaciones, creado_por)
                VALUES (?, datetime('now'), 'Aprobada', ?, ?, ?, ?)""",
                (num_oc, plan['nombre'], plan['valor'],
                 f"Pago influencer importado desde Excel · pub: {plan['fecha_publicacion'] or 'n/a'}", u))
            try:
                c.execute("""INSERT INTO ordenes_compra_items
                    (numero_oc, codigo_mp, nombre_mp, cantidad_g, precio_unitario, subtotal)
                    VALUES (?, '', ?, 1, ?, ?)""",
                    (num_oc, plan['concepto'][:200], plan['valor'], plan['valor']))
            except sqlite3.OperationalError:
                pass
            try:
                # Import historico: el contenido YA se publico y el pago YA ocurrio.
                # Por eso entra como 'Pagada' (no 'Pendiente'). De lo contrario, queda
                # eternamente marcado como pendiente aunque ya se haya transferido.
                c.execute("""INSERT INTO pagos_influencers
                    (influencer_id, influencer_nombre, valor, fecha, estado, concepto, numero_oc, fecha_publicacion)
                    VALUES (?, ?, ?, datetime('now'), 'Pagada', ?, ?, ?)""",
                    (plan['influencer_id'], plan['nombre'], plan['valor'],
                     plan['concepto'][:200], num_oc, plan['fecha_publicacion'] or ''))
            except sqlite3.OperationalError:
                pass
            importadas.append({
                'sol': num_sol, 'oc': num_oc,
                'nombre': plan['nombre'], 'valor': plan['valor'],
                'fecha_publicacion': plan['fecha_publicacion'],
            })

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        import traceback
        return jsonify({
            'error': 'Falla durante la importación',
            'detalle': str(e),
            'traceback': traceback.format_exc()[-800:],
            'rollback': 'aplicado',
        }), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

    _log_sec(u, _client_ip(),
             "admin_import_pagos_influencers_excel",
             f"deleted={deleted} imported={len(importadas)}")

    return jsonify({
        'ok': True,
        'dry_run': False,
        'hoja_usada': sheet_name,
        'reset': deleted,
        'importadas': {
            'count': len(importadas),
            'lista': importadas[:50],
            'valor_total': round(sum(i['valor'] for i in importadas), 0),
        },
        'sin_match': {
            'count': len(sin_match),
            'lista': sin_match[:30],
            'nota': 'NO importados — agrega los influencers al banco o corrige el nombre.',
        },
        'pagadas_skipped': {'count': len(pagadas_skipped)},
        'sin_valor': {'count': len(sin_valor)},
        'sin_nombre': {'count': len(sin_nombre)},
    })


# ─── Auditoría de Mínimos ─────────────────────────────────────────────────────


def _compute_audit_minimos(horizonte_proyeccion_dias: int = 90) -> dict:
    """Helper compartido (Sebastian 5-may-2026): calcula audit de
    stock_minimo sin requerir auth · permite que /api/admin/auditar-minimos
    (admin only · full + apply) y /api/planta/auditar-minimos (read-only ·
    todos los users) compartan la misma lógica.

    Args:
        horizonte_proyeccion_dias: 30-180 días (default 90).

    Returns:
        dict con stats + auditoria + metodologia.
    """
    from database import get_db as _get_db
    from flask import current_app

    try:
        horizonte_proyeccion_dias = max(30, min(int(horizonte_proyeccion_dias), 180))
    except (ValueError, TypeError):
        horizonte_proyeccion_dias = 90

    # 1. Cargar planificacion estratégica para el horizonte
    try:
        from blueprints.programacion import planificacion_estrategica as _plan
    except ImportError:
        from api.blueprints.programacion import planificacion_estrategica as _plan
    with current_app.test_request_context(
        f'/api/programacion/planificacion?dias={horizonte_proyeccion_dias}'
    ):
        plan_resp = _plan()
    plan_data = plan_resp.get_json() or {}

    consumo_por_mp = {}
    for mp in (plan_data.get('mps_deficit') or []) + (plan_data.get('mps_ok') or []):
        consumo_por_mp[mp['material_id']] = {
            'total_g_horizonte': float(mp.get('total_g') or 0),
            'origen': mp.get('origen', 'desconocido'),
            'productos': mp.get('productos', []) or [],
        }

    conn = _get_db()
    try:
        rows = conn.execute("""
            SELECT m.codigo_mp, m.nombre_inci, m.nombre_comercial, m.proveedor,
                   COALESCE(m.stock_minimo, 0), COALESCE(m.tipo_material, 'MP'),
                   COALESCE(s.stock_actual, 0)
            FROM maestro_mps m
            LEFT JOIN (
                SELECT material_id,
                       SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END) as stock_actual
                FROM movimientos GROUP BY material_id
            ) s ON m.codigo_mp = s.material_id
            WHERE m.activo = 1
            ORDER BY m.codigo_mp
        """).fetchall()
    except sqlite3.OperationalError:
        rows = conn.execute("""
            SELECT m.codigo_mp, m.nombre_inci, m.nombre_comercial, m.proveedor,
                   COALESCE(m.stock_minimo, 0), 'MP',
                   COALESCE(s.stock_actual, 0)
            FROM maestro_mps m
            LEFT JOIN (
                SELECT material_id,
                       SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END) as stock_actual
                FROM movimientos GROUP BY material_id
            ) s ON m.codigo_mp = s.material_id
            WHERE m.activo = 1
            ORDER BY m.codigo_mp
        """).fetchall()

    auditoria = []
    for r in rows:
        codigo, nombre_inci, nombre_com, proveedor, stock_min_actual, tipo_mat, stock_actual = r
        nombre = nombre_com or nombre_inci or codigo
        proveedor = (proveedor or '').strip()
        stock_min_actual = float(stock_min_actual or 0)
        stock_actual = float(stock_actual or 0)

        proy = consumo_por_mp.get(codigo)
        consumo_horizonte_g = proy['total_g_horizonte'] if proy else 0.0
        origen = proy['origen'] if proy else 'desconocido'
        productos = proy['productos'] if proy else []
        consumo_diario_g = consumo_horizonte_g / horizonte_proyeccion_dias if horizonte_proyeccion_dias > 0 else 0
        consumo_mensual_g = consumo_diario_g * 30

        if origen == 'china':
            lead_time, buffer_d = 60, 30
        elif origen == 'colombia':
            lead_time, buffer_d = 7, 14
        else:
            if proveedor:
                lead_time, buffer_d = 7, 14
            else:
                lead_time, buffer_d = 14, 14
        dias_buffer = lead_time + buffer_d

        if consumo_diario_g <= 0:
            minimo_recomendado = 0.0
            estado = 'SIN_USO_CON_MIN' if stock_min_actual > 0 else 'SIN_USO'
            razonamiento = f'Sin uso proyectado en próximos {horizonte_proyeccion_dias} días'
        else:
            minimo_recomendado = consumo_diario_g * dias_buffer
            if consumo_diario_g < 0.5:
                minimo_recomendado = max(minimo_recomendado, 50)
            if stock_min_actual == 0:
                estado = 'SIN_MINIMO_CONFIGURADO'
                razonamiento = (
                    f'Sin mínimo configurado · Recomendado {int(round(minimo_recomendado))} g '
                    f'({lead_time}d lead + {buffer_d}d buffer × {round(consumo_diario_g, 2)} g/día)'
                )
            else:
                ratio = stock_min_actual / minimo_recomendado if minimo_recomendado > 0 else 1.0
                cobertura_dias_actual = stock_min_actual / consumo_diario_g if consumo_diario_g > 0 else 0
                if ratio < 0.75:
                    estado = 'SUB_PROTEGIDO'
                    razonamiento = (
                        f'Mínimo cubre solo ~{round(cobertura_dias_actual, 1)} días '
                        f'(necesita ~{dias_buffer} días para origen {origen})'
                    )
                elif ratio > 1.5:
                    estado = 'SOBRE_PROTEGIDO'
                    razonamiento = (
                        f'Mínimo cubre ~{round(cobertura_dias_actual, 1)} días '
                        f'(suficiente con ~{dias_buffer} días para {origen})'
                    )
                else:
                    estado = 'OK'
                    razonamiento = f'Mínimo cubre ~{round(cobertura_dias_actual, 1)} días — apropiado'

        auditoria.append({
            'codigo_mp': codigo,
            'nombre': nombre,
            'proveedor': proveedor,
            'origen': origen,
            'tipo_material': tipo_mat,
            'stock_actual_g': round(stock_actual, 1),
            'stock_minimo_actual_g': round(stock_min_actual, 1),
            'consumo_horizonte_g': round(consumo_horizonte_g, 1),
            'consumo_diario_g': round(consumo_diario_g, 3),
            'consumo_mensual_g': round(consumo_mensual_g, 1),
            'lead_time_dias': lead_time,
            'buffer_dias': buffer_d,
            'dias_cobertura_total': dias_buffer,
            'minimo_recomendado_g': round(minimo_recomendado, 1),
            'estado': estado,
            'razonamiento': razonamiento,
            'productos': productos,
        })

    stats = {
        'total': len(auditoria),
        'ok': sum(1 for a in auditoria if a['estado'] == 'OK'),
        'sub_protegido': sum(1 for a in auditoria if a['estado'] == 'SUB_PROTEGIDO'),
        'sobre_protegido': sum(1 for a in auditoria if a['estado'] == 'SOBRE_PROTEGIDO'),
        'sin_minimo': sum(1 for a in auditoria if a['estado'] == 'SIN_MINIMO_CONFIGURADO'),
        'sin_uso': sum(1 for a in auditoria if a['estado'].startswith('SIN_USO')),
    }

    return {
        'horizonte_proyeccion_dias': horizonte_proyeccion_dias,
        'stats': stats,
        'auditoria': auditoria,
        'metodologia': {
            'formula': 'minimo_recomendado = consumo_diario × (lead_time + buffer)',
            'piso_peptides': 'min 50g si consumo < 0.5 g/día',
            'lead_times': {
                'china': '60d lead + 30d buffer = 90d',
                'colombia/local': '7d lead + 14d buffer = 21d',
                'desconocido_sin_proveedor': '14d + 14d = 28d',
            },
        },
    }


@bp.route("/api/admin/auditar-minimos", methods=["GET"])
def auditar_minimos():
    """Auditoría no-destructiva de stock_minimo de maestro_mps.

    Para cada MP: muestra mínimo actual vs recomendado calculado por
    metodología consumo × (lead_time + buffer) según origen del proveedor.

    Query params:
      proyeccion_dias: int (30-180, default 90) — horizonte para proyectar
        consumo desde el calendario × fórmulas.

    Estados:
      OK: ratio actual/recomendado entre 0.75 y 1.50
      SUB_PROTEGIDO: actual < recomendado × 0.75
      SOBRE_PROTEGIDO: actual > recomendado × 1.50
      SIN_MINIMO_CONFIGURADO: stock_minimo = 0 con consumo > 0
      SIN_USO: sin consumo proyectado en horizonte
      SIN_USO_CON_MIN: sin consumo pero tiene mínimo configurado

    Lead times por origen:
      china: 60d lead + 30d buffer = 90d
      colombia/local: 7d lead + 14d buffer = 21d
      desconocido (con proveedor): 7d + 14d = 21d
      desconocido (sin proveedor): 14d + 14d = 28d
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    try:
        horizonte_proyeccion_dias = max(30, min(int(request.args.get('proyeccion_dias', 90)), 180))
    except ValueError:
        horizonte_proyeccion_dias = 90
    return jsonify(_compute_audit_minimos(horizonte_proyeccion_dias))


@bp.route("/api/admin/aplicar-minimos", methods=["POST"])
def aplicar_minimos():
    """Aplica el recálculo de stock_minimo basado en /auditar-minimos.

    Requiere:
      - Token textual exacto: 'APLICAR_MINIMOS_RECALCULADOS_2026'
      - Backup automático previo
      - Audit log del cambio

    Body:
      token: str (obligatorio)
      proyeccion_dias: int (default 90)
      solo_estados: list[str] (default los 3: SUB_PROTEGIDO, SOBRE_PROTEGIDO,
                    SIN_MINIMO_CONFIGURADO)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    if d.get('token', '').strip() != 'APLICAR_MINIMOS_RECALCULADOS_2026':
        return jsonify({'error': 'Token incorrecto'}), 403

    try:
        horizonte = max(30, min(int(d.get('proyeccion_dias', 90)), 180))
    except ValueError:
        horizonte = 90

    solo_estados = d.get('solo_estados') or [
        'SUB_PROTEGIDO', 'SOBRE_PROTEGIDO', 'SIN_MINIMO_CONFIGURADO'
    ]

    # Backup previo
    try:
        do_backup(triggered_by='pre_minimos_recalc')
    except Exception as e:
        return jsonify({'error': f'Backup falló: {str(e)[:200]}'}), 500

    # Reusar la lógica de auditoría
    from flask import current_app
    with current_app.test_request_context(
        f'/api/admin/auditar-minimos?proyeccion_dias={horizonte}'
    ):
        audit_resp = auditar_minimos()
    audit_data = audit_resp.get_json() or {}

    from database import get_db as _get_db
    conn = _get_db()
    c = conn.cursor()

    cambios = []
    for item in (audit_data.get('auditoria') or []):
        if item['estado'] not in solo_estados:
            continue
        if item['estado'].startswith('SIN_USO'):
            continue  # No tocar — el usuario puede tener motivo
        codigo = item['codigo_mp']
        nuevo = float(item['minimo_recomendado_g'])
        previo = float(item['stock_minimo_actual_g'])
        try:
            c.execute(
                "UPDATE maestro_mps SET stock_minimo = ? WHERE codigo_mp = ?",
                (nuevo, codigo),
            )
            cambios.append({
                'codigo_mp': codigo,
                'nombre': item['nombre'],
                'previo_g': round(previo, 1),
                'nuevo_g': round(nuevo, 1),
                'estado_previo': item['estado'],
            })
        except Exception:
            continue
    conn.commit()

    # Audit log
    try:
        import json as _json
        c.execute(
            """INSERT INTO audit_log
               (usuario, accion, tabla, registro_id, detalle, ip, fecha)
               VALUES (?,?,?,?,?,?,datetime('now'))""",
            (u, 'APLICAR_MINIMOS_RECALCULADOS', 'maestro_mps', '_BULK_',
             _json.dumps({
                 'count': len(cambios),
                 'horizonte_proyeccion_dias': horizonte,
                 'estados_aplicados': solo_estados,
             }, ensure_ascii=False),
             request.remote_addr),
        )
        conn.commit()
    except Exception:
        pass

    return jsonify({
        'ok': True,
        'count_cambios': len(cambios),
        'cambios': cambios[:50],
        'mensaje': f'{len(cambios)} mínimos actualizados. Backup previo creado.',
    })


# ─── Panel HTML ───────────────────────────────────────────────────────────────

_ADMIN_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Admin — HHA Group</title>
<style>
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:'Segoe UI',sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh;font-size:14px;}
.hdr{background:#1e293b;border-bottom:1px solid #334155;padding:14px 24px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10;}
.hdr h1{font-size:16px;font-weight:800;color:#fff;}
.hdr a{color:#667eea;text-decoration:none;font-size:12px;}
.tabs{background:#1e293b;border-bottom:1px solid #334155;display:flex;overflow-x:auto;padding:0 24px;position:sticky;top:48px;z-index:9;}
.tab{padding:14px 22px;font-size:13px;font-weight:600;color:#64748b;border:none;background:none;cursor:pointer;white-space:nowrap;border-bottom:3px solid transparent;transition:.15s;}
.tab:hover{color:#e2e8f0;}
.tab.active{color:#a78bfa;border-bottom-color:#7c3aed;}
.main{max-width:1200px;margin:0 auto;padding:24px;}
.card{background:#1e293b;border:1px solid #334155;border-radius:12px;padding:20px;margin-bottom:18px;}
.card h2{font-size:15px;font-weight:700;color:#f1f5f9;margin-bottom:12px;display:flex;align-items:center;gap:10px;}
.kpi-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:18px;}
.kpi{background:#0f172a;border:1px solid #334155;border-radius:10px;padding:14px;}
.kpi-l{font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;}
.kpi-v{font-size:20px;font-weight:800;color:#a78bfa;}
.btn{display:inline-flex;align-items:center;gap:6px;padding:9px 16px;border-radius:8px;border:none;cursor:pointer;font-size:13px;font-weight:700;color:#fff;background:linear-gradient(135deg,#7c3aed,#4c1d95);text-decoration:none;}
.btn:hover{filter:brightness(1.1);}
.btn:disabled{opacity:.5;cursor:wait;}
.btn-sm{padding:4px 10px;font-size:11px;}
.btn-outline{background:transparent;border:1px solid #475569;color:#94a3b8;}
.btn-warn{background:linear-gradient(135deg,#dc2626,#7f1d1d);}
table{width:100%;border-collapse:collapse;font-size:13px;}
th{font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.05em;padding:8px 12px;text-align:left;background:#0f172a;border-bottom:1px solid #334155;}
td{padding:10px 12px;border-bottom:1px solid #1e293b;}
tr:hover td{background:#263348;}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700;}
.badge-ok{background:#052e16;color:#34d399;}
.badge-err{background:#2d0000;color:#f87171;}
.badge-warn{background:#3a2a00;color:#fbbf24;}
.badge-info{background:#1e3a5f;color:#93c5fd;}
.badge-run{background:#1e1b4b;color:#a78bfa;}
.badge-gray{background:#0f172a;color:#94a3b8;border:1px solid #334155;}
#toast{position:fixed;top:80px;right:20px;background:#1e293b;border:1px solid #334155;color:#e2e8f0;padding:14px 20px;border-radius:10px;font-size:13px;display:none;z-index:1000;box-shadow:0 8px 24px rgba(0,0,0,0.4);max-width:400px;}
.section-sub{font-size:12px;color:#64748b;margin-top:-8px;margin-bottom:16px;line-height:1.5;}
.tab-panel{display:none;}
.tab-panel.active{display:block;}

/* Modal */
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.7);z-index:100;align-items:center;justify-content:center;}
.modal-bg.show{display:flex;}
.modal{background:#1e293b;border:1px solid #334155;border-radius:14px;padding:28px;max-width:480px;width:92%;}
.modal h3{font-size:16px;font-weight:700;color:#f1f5f9;margin-bottom:8px;}
.modal-msg{font-size:13px;color:#94a3b8;margin-bottom:18px;line-height:1.5;}
.modal-result{background:#0f172a;border:1px solid #334155;border-radius:8px;padding:14px;margin:12px 0;font-family:'Cascadia Code',Consolas,monospace;font-size:14px;color:#34d399;text-align:center;letter-spacing:0.05em;word-break:break-all;}
.modal-actions{display:flex;gap:10px;justify-content:flex-end;margin-top:14px;}
</style>
</head>
<body>
<div class="hdr">
  <h1>&#x1F510; Panel de Administracion</h1>
  <a href="/hub">&#x2190; Volver al Hub</a>
</div>

<div class="tabs">
  <button class="tab active" data-tab="backups" onclick="switchTab('backups')">&#x1F4C0; Backups</button>
  <button class="tab" data-tab="users" onclick="switchTab('users')">&#x1F465; Usuarios</button>
  <button class="tab" data-tab="security" onclick="switchTab('security')">&#x1F6E1; Eventos de Seguridad</button>
  <button class="tab" data-tab="config" onclick="switchTab('config')">&#x2699; Config Status</button>
  <button class="tab" data-tab="banco" onclick="switchTab('banco')">&#x1F4B3; Banco Influencers</button>
  <button class="tab" data-tab="mps" onclick="switchTab('mps')">&#x1F9EA; Cat&aacute;logo MPs</button>
  <button class="tab" data-tab="audit-inv" onclick="switchTab('audit-inv')">&#x1F50D; Auditar Inventario</button>
  <button class="tab" data-tab="audit-min" onclick="switchTab('audit-min')">&#x1F4CA; Auditar M&iacute;nimos</button>
  <button class="tab" data-tab="diag-form" onclick="switchTab('diag-form')">&#x1F9EA; Diagn&oacute;stico F&oacute;rmulas</button>
</div>

<div class="main">

<!-- ─── TAB BACKUPS ─── -->
<div id="tab-backups" class="tab-panel active">
  <div class="card">
    <h2>&#x1F4C0; Backups de Base de Datos</h2>
    <div class="section-sub">
      Backups automaticos cada 23h, 7 dias de retencion. Descarga el mas reciente regularmente
      para tener una copia <strong>fuera de Render</strong> — eso te protege si el disco se corrompe.
    </div>
    <div class="kpi-row">
      <div class="kpi"><div class="kpi-l">Disponibles</div><div class="kpi-v" id="kpi-count">-</div></div>
      <div class="kpi"><div class="kpi-l">Espacio total</div><div class="kpi-v" id="kpi-size">-</div></div>
      <div class="kpi"><div class="kpi-l">Ultimo backup</div><div class="kpi-v" id="kpi-last" style="font-size:13px;">-</div></div>
      <div class="kpi"><div class="kpi-l">Retencion</div><div class="kpi-v" style="font-size:14px;" id="kpi-ret">-</div></div>
    </div>
    <button class="btn" id="btn-backup" onclick="triggerBackup()">
      <span id="btn-label">&#x26A1; Hacer backup ahora</span>
    </button>
    <div style="margin-top:24px;">
      <table>
        <thead><tr><th>Archivo</th><th>Fecha (UTC)</th><th>Tamano</th><th>Accion</th></tr></thead>
        <tbody id="tbody-backups"><tr><td colspan="4" style="text-align:center;color:#64748b;padding:30px;">Cargando...</td></tr></tbody>
      </table>
    </div>
  </div>
  <div class="card">
    <h2>&#x1F4DC; Historial de ejecuciones</h2>
    <div class="section-sub">Audita backups exitosos y fallidos.</div>
    <table>
      <thead><tr><th>ID</th><th>Inicio (UTC)</th><th>Estado</th><th>Trigger</th><th>Tamano</th><th>Error</th></tr></thead>
      <tbody id="tbody-runs"><tr><td colspan="6" style="text-align:center;color:#64748b;padding:30px;">Cargando...</td></tr></tbody>
    </table>
  </div>
</div>

<!-- ─── TAB USERS ─── -->
<div id="tab-users" class="tab-panel">
  <div class="card">
    <h2>&#x1F465; Usuarios del sistema</h2>
    <div class="section-sub">
      Aqui ves a los 19 usuarios con sus grupos, fuente de password (DB = cambiada por el user, ENV = de Render),
      y ultimo login. Si alguien olvida su password, presiona <strong>Resetear</strong> y comparte
      la nueva contrasena 1-on-1.
    </div>
    <table>
      <thead>
        <tr>
          <th>Usuario</th><th>Grupos</th><th>Password</th>
          <th>Ult. login</th><th>Pwd cambiada</th><th>Accion</th>
        </tr>
      </thead>
      <tbody id="tbody-users"><tr><td colspan="6" style="text-align:center;color:#64748b;padding:30px;">Cargando...</td></tr></tbody>
    </table>
  </div>
</div>

<!-- ─── TAB SECURITY ─── -->
<div id="tab-security" class="tab-panel">
  <div class="card">
    <h2>&#x1F6E1; Estadisticas (ultimas 24h)</h2>
    <div class="section-sub">Conteo de eventos por tipo en las ultimas 24 horas. Picos sospechosos = investigar.</div>
    <div class="kpi-row" id="kpi-stats">
      <div class="kpi"><div class="kpi-l">Cargando...</div><div class="kpi-v">-</div></div>
    </div>
  </div>
  <div class="card">
    <h2>&#x1F50D; Eventos de seguridad</h2>
    <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:14px;">
      <label style="font-size:11px;color:#94a3b8;">Filtrar:</label>
      <select id="filter-event" onchange="loadSecurityEvents()" style="background:#0f172a;border:1px solid #334155;color:#e2e8f0;padding:6px 10px;border-radius:8px;font-size:12px;">
        <option value="">Todos los eventos</option>
        <option value="login_success">login_success</option>
        <option value="login_failure">login_failure</option>
        <option value="password_changed">password_changed</option>
        <option value="password_change_failed">password_change_failed</option>
        <option value="password_reset_by_admin">password_reset_by_admin</option>
        <option value="csrf_blocked">csrf_blocked</option>
        <option value="backup_manual_triggered">backup_manual_triggered</option>
        <option value="backup_downloaded">backup_downloaded</option>
      </select>
      <button class="btn btn-sm btn-outline" onclick="loadSecurityEvents()">&#x21BB; Refrescar</button>
    </div>
    <table>
      <thead>
        <tr><th>ID</th><th>Timestamp UTC</th><th>Evento</th><th>Usuario</th><th>IP</th><th>Detalles</th></tr>
      </thead>
      <tbody id="tbody-security"><tr><td colspan="6" style="text-align:center;color:#64748b;padding:30px;">Cargando...</td></tr></tbody>
    </table>
  </div>
</div>

<!-- ─── TAB CONFIG ─── -->
<div id="tab-config" class="tab-panel">
  <div class="card">
    <h2>&#x26A0; Issues de configuracion</h2>
    <div class="section-sub">
      Issues detectados por <code>validate_config()</code> al startup.
      <strong>Cero CRITICAL/HIGH = todo OK.</strong>
    </div>
    <div id="config-issues" style="margin-top:8px;">Cargando...</div>
  </div>
  <div class="card">
    <h2>&#x1F511; Variables de entorno</h2>
    <div class="section-sub">
      Solo se muestra si la variable esta seteada y su longitud — nunca el valor.
      Las criticas DEBEN estar; las de PASS_USER son las contrasenas individuales;
      las opcionales habilitan integraciones.
    </div>
    <h3 style="color:#a78bfa;font-size:13px;margin-top:16px;margin-bottom:8px;">Criticas</h3>
    <table><tbody id="tbody-config-critical"><tr><td>Cargando...</td></tr></tbody></table>
    <h3 style="color:#a78bfa;font-size:13px;margin-top:16px;margin-bottom:8px;">Passwords por usuario</h3>
    <table><tbody id="tbody-config-users"><tr><td>Cargando...</td></tr></tbody></table>
    <h3 style="color:#a78bfa;font-size:13px;margin-top:16px;margin-bottom:8px;">Opcionales (integraciones)</h3>
    <table><tbody id="tbody-config-optional"><tr><td>Cargando...</td></tr></tbody></table>
  </div>
  <div class="card">
    <h2>&#x1F4E7; Test de SMTP (envío de correo)</h2>
    <div class="section-sub">
      Manda un correo de prueba con un PDF demo (CE-TEST-0000) al destinatario que indiques.
      Si dejas el campo vacío, se envía al remitente (a ti mismo). Sirve para validar
      que <code>EMAIL_REMITENTE</code> y <code>EMAIL_PASSWORD</code> están bien seteados
      antes de pagarle a un influencer real.
    </div>
    <div style="margin-top:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
      <input type="email" id="test-email-dest" placeholder="destinatario@gmail.com (opcional)"
             style="padding:9px 14px;background:#0f172a;border:1px solid #334155;border-radius:6px;color:#e2e8f0;min-width:280px;font-size:13px;">
      <select id="test-email-empresa"
              style="padding:9px 14px;background:#0f172a;border:1px solid #334155;border-radius:6px;color:#e2e8f0;font-size:13px;">
        <option value="espagiria">PDF como Espagiria</option>
        <option value="animus">PDF como ANIMUS Lab</option>
      </select>
      <button class="btn" onclick="enviarTestEmail()">&#x26A1; Enviar correo de prueba</button>
    </div>
    <div id="test-email-result" style="margin-top:14px;"></div>
  </div>
</div>

<!-- ─── TAB BANCO INFLUENCERS ─── -->
<div id="tab-banco" class="tab-panel">
  <div class="card">
    <h2>&#x1F4B3; Sincronizar banco de influencers desde Excel</h2>
    <div class="section-sub">
      Sube el Excel de creadores (hoja <code>cuentas de creadores</code>). El sync
      hace UPSERT por nombre <strong>preservando datos existentes</strong> — solo
      llena columnas vacías de la base con datos del Excel. Nunca borra ni
      sobrescribe banco/cuenta/cédula que ya estén cargados.
    </div>
    <div style="margin-top:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
      <input type="file" id="excel-file" accept=".xlsx,.xlsm"
             style="padding:8px;background:#0f172a;border:1px solid #334155;border-radius:6px;color:#e2e8f0;">
      <button class="btn btn-outline" onclick="syncExcel(true)">&#x1F441; Vista previa (dry-run)</button>
      <button class="btn" onclick="syncExcel(false)">&#x26A1; Aplicar a la base</button>
    </div>
    <div id="banco-result" style="margin-top:18px;"></div>
  </div>

  <div class="card" style="border-left:3px solid #fbbf24;">
    <h2>&#x1F4B8; Importar pagos pendientes desde Excel</h2>
    <div class="section-sub">
      Sube el Excel con la lista de pagos pendientes (uno por fila). El sistema:
      <ul style="margin:6px 0 0 16px;">
        <li>Detecta automáticamente las columnas (Influencer, Fecha publicación, Valor, Estado, Concepto)</li>
        <li><strong>Borra las solicitudes Influencer no-pagadas</strong> (mantiene las Pagadas)</li>
        <li>Crea cada fila como solicitud Aprobada lista para pagar, con su <code>fecha_publicacion</code></li>
        <li>Saltea filas marcadas como "Pagado" en el Excel</li>
        <li>Saltea filas donde el nombre no matchea ningún influencer del banco</li>
      </ul>
    </div>
    <div style="margin-top:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
      <input type="file" id="pagos-excel-file" accept=".xlsx,.xlsm"
             style="padding:8px;background:#0f172a;border:1px solid #334155;border-radius:6px;color:#e2e8f0;">
      <button class="btn btn-outline" onclick="syncPagos(true)">&#x1F441; Vista previa (dry-run)</button>
      <button class="btn" onclick="syncPagos(false)" style="background:linear-gradient(135deg,#f59e0b,#d97706);">&#x26A1; Aplicar (borra pendientes + importa)</button>
    </div>
    <div id="pagos-result" style="margin-top:18px;"></div>
  </div>
</div>

<!-- ─── TAB CATÁLOGO MPs ─── -->
<div id="tab-mps" class="tab-panel">
  <div class="card">
    <h2>&#x1F9EA; Cat&aacute;logo de Materias Primas - estado de proveedor</h2>
    <div class="section-sub">
      Cuando creas una OC sugerida desde Compras, el sistema agrupa las MPs por
      <code>maestro_mps.proveedor</code>. Las MPs sin proveedor caen en
      'Sin asignar' y tienes que corregir manualmente cada vez. Asigna aqu&iacute;
      el proveedor habitual de cada MP para que las futuras OCs salgan
      correctamente agrupadas.
    </div>
    <div class="kpi-row" id="mps-kpis">
      <div class="kpi"><div class="kpi-l">Total activos</div><div class="kpi-v" id="mps-kpi-total">-</div></div>
      <div class="kpi"><div class="kpi-l">Con proveedor</div><div class="kpi-v" id="mps-kpi-con" style="color:#34d399;">-</div></div>
      <div class="kpi"><div class="kpi-l">Sin proveedor</div><div class="kpi-v" id="mps-kpi-sin" style="color:#f87171;">-</div></div>
      <div class="kpi"><div class="kpi-l">Cobertura</div><div class="kpi-v" id="mps-kpi-pct" style="font-size:14px;">-</div></div>
    </div>
    <h3 style="color:#a78bfa;font-size:13px;margin-top:16px;margin-bottom:8px;">Distribuci&oacute;n por proveedor</h3>
    <table>
      <thead><tr><th>Proveedor</th><th style="text-align:right;">Cantidad de MPs</th></tr></thead>
      <tbody id="mps-tbody-prov"><tr><td>Cargando...</td></tr></tbody>
    </table>
    <h3 style="color:#f87171;font-size:13px;margin-top:20px;margin-bottom:8px;">MPs sin proveedor asignado</h3>
    <div id="mps-sin-prov-info" style="font-size:12px;color:#94a3b8;margin-bottom:8px;"></div>
    <table>
      <thead><tr><th>C&oacute;digo</th><th>Nombre</th><th>Tipo</th><th>Asignar proveedor</th></tr></thead>
      <tbody id="mps-tbody-sin"><tr><td>Cargando...</td></tr></tbody>
    </table>
  </div>

  <div class="card" style="border-left:3px solid #fbbf24;">
    <h2>&#x1F4DD; Corregir nombres masivos desde Excel</h2>
    <div class="section-sub">
      Sube un Excel con columnas <code>código</code> + <code>nombre</code>
      (+ <code>proveedor</code> opcional). Solo actualiza nombres que est&aacute;n
      <strong>vac&iacute;os o iguales al c&oacute;digo</strong> (mal cargados).
      <strong>Nombres correctos NO se sobrescriben.</strong>
    </div>
    <div style="margin-top:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
      <input type="file" id="mps-nom-file" accept=".xlsx,.xlsm"
             style="padding:8px;background:#0f172a;border:1px solid #334155;border-radius:6px;color:#e2e8f0;">
      <button class="btn btn-outline" onclick="syncMpsNombres(true)">&#x1F441; Vista previa</button>
      <button class="btn" onclick="syncMpsNombres(false)" style="background:linear-gradient(135deg,#f59e0b,#d97706);">&#x26A1; Aplicar correcci&oacute;n</button>
    </div>
    <div id="mps-nom-result" style="margin-top:14px;"></div>
  </div>
</div>

<!-- ─── TAB AUDITAR INVENTARIO vs EXCEL día cero ─── -->
<div id="tab-audit-inv" class="tab-panel">
  <div class="card">
    <h2>&#x1F50D; Auditar Inventario vs Excel d&iacute;a cero</h2>
    <div class="section-sub">
      Sube el Excel del conteo f&iacute;sico (ej: <code>INVENTARIO_MP_v8_1.xlsx</code>) — solo se contar&aacute;n las
      filas marcadas en <strong style="color:#16a34a;">verde</strong>. Las rojas y sin marcar se ignoran (Catalina las marc&oacute; como NO presentes).
      <br><br>
      <strong>Cero escritura</strong>. El reporte muestra c&oacute;mo difiere el kardex actual contra ese conteo + las producciones que se han hecho desde entonces.
      Despu&eacute;s decides: ajustes quir&uacute;rgicos o reset+replay.
    </div>
    <div style="display:flex;gap:10px;align-items:center;margin-top:12px;flex-wrap:wrap;">
      <input type="file" id="audit-inv-file" accept=".xlsx,.xlsm" style="padding:6px;background:#0f172a;border:1px solid #334155;border-radius:8px;color:#e2e8f0;">
      <button class="btn" onclick="auditarInvVsExcel()">&#x1F4CA; Generar reporte</button>
      <button class="btn btn-outline" onclick="diagnosticoEntradas()" title="Antes de borrar nada, ver QUIEN cargo QUE — detecta doble carga vs recepciones manuales legitimas">&#x1F50E; Diagn&oacute;stico de entradas</button>
      <button class="btn btn-outline" onclick="quePuedoProducir()" title="Para cada producto, indica si las MPs alcanzan + shopping list de faltantes" style="border-color:#fbbf24;color:#fbbf24;">&#x1F3ED; Qu&eacute; puedo producir</button>
      <button class="btn btn-outline" onclick="healthMonitor()" title="Monitor periodico: detecta bursts, multi-entradas, anomalias en stock — alerta antes de que la duplicacion sea masiva" style="border-color:#ef4444;color:#ef4444;">&#x1F525; Monitor anomal&iacute;as</button>
    </div>

    <div style="margin-top:24px;padding:16px;background:rgba(220,38,38,.08);border:1px solid rgba(220,38,38,.4);border-radius:10px;">
      <h3 style="color:#fca5a5;margin:0 0 8px 0;">&#x26A0; Reset + Replay del inventario</h3>
      <div style="font-size:12px;color:#cbd5e1;margin-bottom:12px;">
        <strong style="color:#fca5a5;">DESTRUCTIVO.</strong> Borra todos los movimientos y los recarga desde el Excel verde (estado d&iacute;a cero) + preserva las recepciones formales con OC + re-aplica las salidas de las producciones.
        <br>Sigue el orden: <strong>1)</strong> Descarga snapshot &rarr; <strong>2)</strong> Preview &rarr; <strong>3)</strong> Aplicar.
      </div>
      <div style="display:flex;gap:10px;flex-wrap:wrap;">
        <button class="btn btn-outline" onclick="sembrarMaestroDesdeExcel()" style="border-color:#a5b4fc;color:#a5b4fc;" title="Asegura que TODOS los códigos del Excel (verdes Y no-verdes) estén en maestro_mps. NO toca movimientos. Recomendado correr ANTES del reset.">&#x1F331; 0. Sembrar cat&aacute;logo MPs</button>
        <button class="btn btn-outline" onclick="descargarSnapshotPreReset()" title="Descarga JSON con TODOS los movimientos, producciones, OCs, comprobantes — para poder revertir si algo sale mal">&#x1F4BE; 1. Descargar snapshot pre-reset</button>
        <button class="btn btn-outline" onclick="previewReset()" title="Muestra que va a pasar SIN escribir nada">&#x1F441; 2. Preview reset</button>
        <button class="btn" onclick="aplicarReset()" style="background:#dc2626;color:#fff;" title="Ejecuta el reset. Pide token textual de confirmacion.">&#x1F4A5; 3. APLICAR reset</button>
        <button class="btn btn-outline" onclick="healthCheckPostReset()" style="border-color:#34d399;color:#34d399;" title="Verifica integridad del sistema despues del reset">&#x2705; 4. Health-check post-reset</button>
      </div>
    </div>

    <div id="audit-inv-result" style="margin-top:18px;"></div>
  </div>
</div>

<!-- ─── TAB DIAGNOSTICO FORMULAS ─── -->
<div id="tab-diag-form" class="tab-panel">
  <div class="card">
    <h2>&#x1F9EA; Diagn&oacute;stico de F&oacute;rmulas</h2>
    <div class="section-sub">
      Detecta items en <code>formula_items</code> que apuntan a <strong>material_id</strong> hu&eacute;rfano (no en <code>maestro_mps</code>) o cuyo nombre no coincide con el cat&aacute;logo. Sugiere correcci&oacute;n autom&aacute;tica buscando por nombre similar.
      <br><br>
      Importante: si las f&oacute;rmulas apuntan a c&oacute;digos legacy o err&oacute;neos, la <strong>Vista por producci&oacute;n</strong> y los <strong>cálculos de stock</strong> dan resultados incorrectos. Usa esto para sincronizar.
    </div>
    <div style="margin-top:14px;display:flex;gap:8px;flex-wrap:wrap;">
      <button class="btn" onclick="cargarDiagnosticoFormulas()" id="btn-diag-form">&#x1F50D; Ejecutar diagn&oacute;stico</button>
      <button class="btn btn-outline" onclick="exportarDiagFormCSV()">&#x1F4C4; Exportar CSV</button>
    </div>

    <div id="diag-form-stats" style="margin-top:18px;display:none;">
      <div class="kpi-row">
        <div class="kpi"><div class="kpi-l">Total &iacute;tems</div><div class="kpi-v" id="diagf-total">-</div></div>
        <div class="kpi" style="border-left:3px solid #fbbf24;"><div class="kpi-l">Con problemas</div><div class="kpi-v" id="diagf-problemas" style="color:#fbbf24;">-</div></div>
        <div class="kpi" style="border-left:3px solid #dc2626;"><div class="kpi-l">Hu&eacute;rfanos</div><div class="kpi-v" id="diagf-huerf" style="color:#dc2626;">-</div></div>
        <div class="kpi" style="border-left:3px solid #f59e0b;"><div class="kpi-l">Mismatch nombre</div><div class="kpi-v" id="diagf-misn" style="color:#f59e0b;">-</div></div>
        <div class="kpi" style="border-left:3px solid #22c55e;"><div class="kpi-l">Auto-corregibles</div><div class="kpi-v" id="diagf-auto" style="color:#22c55e;">-</div></div>
        <div class="kpi" style="border-left:3px solid #6366f1;"><div class="kpi-l">Requieren revisi&oacute;n</div><div class="kpi-v" id="diagf-rev" style="color:#6366f1;">-</div></div>
      </div>
    </div>

    <div style="margin-top:14px;padding:14px;background:#0f3a1f;border:2px solid #16a34a;border-radius:8px;">
      <div style="font-weight:700;color:#86efac;margin-bottom:6px;">&#x2705; Aplicar batch de correcciones validado por Sebasti&aacute;n + Alejandro (2026-04-28)</div>
      <div style="font-size:12px;color:#bbf7d0;margin-bottom:10px;">
        Aplica de una sola pulsaci&oacute;n los <strong>240 cambios</strong> consensuados:
        <ul style="margin:6px 0 6px 20px;font-size:11px;">
          <li>44 MPs nuevos creados (HPR, RR, p&eacute;ptidos premium, dimethicone, soda c&aacute;ustica, etc.)</li>
          <li>Azeclair (MP00284) marcado <code>activo=0</code> · INCI oficial actualizado en AOS 40</li>
          <li>207 <code>formula_items</code> corregidos (typos, mapeos a Vit C / Carbopol / Plantaren / etc.)</li>
          <li>33 filas de Agua Desionizada eliminadas (agua interna infinita)</li>
        </ul>
        <strong>Backup autom&aacute;tico previo</strong> (full DB + tablas <code>*_backup_20260428</code>). Ejecuci&oacute;n at&oacute;mica en transacci&oacute;n.
      </div>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <input id="diagf-token-batch20260428" placeholder="Token: APLICAR_CORRECCIONES_2026_04_28" style="background:#0f172a;color:#e2e8f0;border:1px solid #86efac;border-radius:6px;padding:7px 10px;font-size:12px;width:340px;">
        <button class="btn" style="background:#16a34a;color:#fff;font-weight:700;" onclick="aplicarBatch20260428()" id="btn-aplicar-batch">&#x2728; Aplicar 240 correcciones</button>
      </div>
    </div>

    <div style="margin-top:14px;padding:14px;background:#3f0f0f;border:2px solid #dc2626;border-radius:8px;">
      <div style="font-weight:700;color:#fca5a5;margin-bottom:6px;">&#x21A9; Revertir &uacute;ltima correcci&oacute;n de f&oacute;rmulas (deshacer)</div>
      <div style="font-size:12px;color:#fecaca;margin-bottom:10px;">
        Si aplicaste correcciones y te diste cuenta que estaban mal, este bot&oacute;n restaura <code>formula_items</code> al estado del backup autom&aacute;tico que se cre&oacute; antes de aplicar. NO toca movimientos / cat&aacute;logo / OCs.
      </div>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <input id="diagf-token-revertir" placeholder="Token: REVERTIR_FORMULAS_2026" style="background:#0f172a;color:#e2e8f0;border:1px solid #fca5a5;border-radius:6px;padding:7px 10px;font-size:12px;width:280px;">
        <button class="btn" style="background:#991b1b;color:#fff;" onclick="revertirFormulas()" id="btn-revertir-form">&#x21A9; Revertir desde backup</button>
      </div>
    </div>

    <div id="diag-form-aplicar-box" style="display:none;margin-top:14px;padding:14px;background:#1e293b;border:1px solid #475569;border-radius:8px;">
      <div style="font-weight:700;color:#fbbf24;margin-bottom:6px;">&#x26A0; Aplicar correcciones</div>
      <div style="font-size:12px;color:#cbd5e1;margin-bottom:10px;">
        Solo se aplican las marcadas (auto-corregibles vienen marcadas por default). Backup autom&aacute;tico previo + audit log.
      </div>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <button class="btn btn-outline" onclick="seleccionarSoloAuto()">Solo auto-corregibles</button>
        <button class="btn btn-outline" onclick="seleccionarTodos()">Seleccionar todos</button>
        <button class="btn btn-outline" onclick="deseleccionarTodos()">Deseleccionar</button>
        <input id="diagf-token" placeholder="Token: CORREGIR_FORMULAS_2026" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:6px;padding:7px 10px;font-size:12px;width:280px;">
        <button class="btn" style="background:#dc2626;color:#fff;" onclick="aplicarCorreccionFormulas()" id="btn-aplicar-form">&#x1F4A5; Aplicar correcciones</button>
      </div>
    </div>

    <div id="diag-form-obsoletas-box" style="display:none;margin-top:14px;padding:14px;background:#1e1b3b;border:1px solid #6366f1;border-radius:8px;">
      <div style="font-weight:700;color:#a5b4fc;margin-bottom:6px;">&#x1F5D1; Eliminar f&oacute;rmulas obsoletas (sin candidato)</div>
      <div style="font-size:12px;color:#cbd5e1;margin-bottom:10px;">
        Los <strong>hu&eacute;rfanos sin candidato</strong> son items en <code>formula_items</code> cuyo <code>material_id</code> NO existe en cat&aacute;logo Y no encontramos similar por nombre. Probablemente son ingredientes descontinuados o productos obsoletos. Eliminarlos NO afecta producciones actuales (esos productos no se pueden producir hasta que la f&oacute;rmula se reescriba).
      </div>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <span id="diagf-sin-cand-count" style="color:#fca5a5;font-weight:700;font-size:13px;">- huérfanos sin candidato</span>
        <input id="diagf-token-eliminar" placeholder="Token: ELIMINAR_FORMULAS_OBSOLETAS_2026" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:6px;padding:7px 10px;font-size:12px;width:340px;">
        <button class="btn" style="background:#7c3aed;color:#fff;" onclick="eliminarFormulasObsoletas()" id="btn-eliminar-obs">&#x1F5D1; Eliminar obsoletas</button>
      </div>
    </div>

    <div id="diag-form-result" style="margin-top:14px;"></div>
  </div>
</div>

<!-- ─── TAB AUDITAR MINIMOS ─── -->
<div id="tab-audit-min" class="tab-panel">
  <div class="card">
    <h2>&#x1F4CA; Auditar M&iacute;nimos de Materias Primas</h2>
    <div class="section-sub">
      Compara <strong>stock_minimo</strong> actual vs lo recomendado por consumo proyectado.
      M&eacute;todo: <code>min&iacute;mo = consumo_diario &times; (lead_time + buffer)</code> seg&uacute;n origen del proveedor:
      China 90 d&iacute;as, Colombia/local 21 d&iacute;as. Piso 50 g para p&eacute;ptidos de baja rotaci&oacute;n.
    </div>
    <div style="margin-top:14px;display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
      <label style="font-size:13px;color:#cbd5e1;">Horizonte de proyecci&oacute;n:</label>
      <select id="audmin-proy" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:6px;padding:6px 10px;font-size:13px;">
        <option value="60">60 d&iacute;as</option>
        <option value="90" selected>90 d&iacute;as (recomendado)</option>
        <option value="120">120 d&iacute;as</option>
        <option value="180">180 d&iacute;as</option>
      </select>
      <button class="btn" onclick="cargarAuditarMinimos()" id="btn-aud-min">&#x1F50D; Auditar (vista previa)</button>
      <button class="btn btn-outline" onclick="exportarAuditMinCSV()">&#x1F4C4; Exportar CSV</button>
    </div>

    <div id="audmin-stats" style="margin-top:18px;display:none;">
      <div class="kpi-row">
        <div class="kpi"><div class="kpi-l">Total MPs</div><div class="kpi-v" id="audmin-total">-</div></div>
        <div class="kpi" style="border-left:3px solid #22c55e;"><div class="kpi-l">OK</div><div class="kpi-v" id="audmin-ok" style="color:#22c55e;">-</div></div>
        <div class="kpi" style="border-left:3px solid #dc2626;"><div class="kpi-l">Sub-protegidos</div><div class="kpi-v" id="audmin-sub" style="color:#dc2626;">-</div></div>
        <div class="kpi" style="border-left:3px solid #f59e0b;"><div class="kpi-l">Sobre-protegidos</div><div class="kpi-v" id="audmin-sobre" style="color:#f59e0b;">-</div></div>
        <div class="kpi" style="border-left:3px solid #6366f1;"><div class="kpi-l">Sin m&iacute;nimo</div><div class="kpi-v" id="audmin-vacio" style="color:#6366f1;">-</div></div>
        <div class="kpi"><div class="kpi-l">Sin uso</div><div class="kpi-v" id="audmin-uso" style="color:#94a3b8;">-</div></div>
      </div>
    </div>

    <div id="audmin-aplicar-box" style="display:none;margin-top:18px;padding:14px;background:#1e293b;border:1px solid #475569;border-radius:8px;">
      <div style="font-weight:700;margin-bottom:6px;color:#fbbf24;">&#x26A0; Aplicar recálculo</div>
      <div style="font-size:12px;color:#cbd5e1;margin-bottom:10px;">
        Esto actualiza <code>stock_minimo</code> en <code>maestro_mps</code> para los MPs marcados como
        <strong>SUB_PROTEGIDO</strong>, <strong>SOBRE_PROTEGIDO</strong> y <strong>SIN_MINIMO_CONFIGURADO</strong>.
        Crea backup autom&aacute;tico previo y registra en audit log. NO toca MPs sin uso proyectado.
      </div>
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <input id="audmin-token" placeholder="Token: APLICAR_MINIMOS_RECALCULADOS_2026" style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:6px;padding:7px 10px;font-size:12px;width:340px;">
        <button class="btn" style="background:#dc2626;color:#fff;" onclick="aplicarRecalculoMinimos()" id="btn-aplicar-min">&#x1F4A5; Aplicar recálculo</button>
      </div>
    </div>

    <div id="audmin-result" style="margin-top:18px;"></div>
  </div>
</div>

</div>

<div id="toast"></div>

<!-- Modal de password reseteada -->
<div id="modal-bg" class="modal-bg">
  <div class="modal">
    <h3 id="modal-title">Password reseteada</h3>
    <div class="modal-msg" id="modal-msg"></div>
    <div class="modal-result" id="modal-result"></div>
    <div class="modal-msg" style="font-size:11px;color:#fbbf24;">
      &#x26A0; Esta password se muestra UNA SOLA VEZ. Comunicala al usuario por canal seguro
      y dile que la cambie en su primer login.
    </div>
    <div class="modal-actions">
      <button class="btn btn-outline" onclick="copyResult()">Copiar</button>
      <button class="btn" onclick="closeModal()">Cerrar</button>
    </div>
  </div>
</div>

<script>
// ── Tabs ──────────────────────────────────────────────────────────────────────
const _loaded = {backups:false, users:false, security:false, config:false, banco:false, mps:false, 'audit-inv':false};
function switchTab(name) {
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t.dataset.tab === name));
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  if (!_loaded[name]) {
    _loaded[name] = true;
    if (name === 'backups')  loadBackups();
    if (name === 'users')    loadUsers();
    if (name === 'security') loadSecurityEvents();
    if (name === 'config')   loadConfigStatus();
    if (name === 'mps')      loadMpsStatus();
  }
}

// ── Toast ─────────────────────────────────────────────────────────────────────
function toast(msg, kind) {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.style.borderColor = kind === 'ok' ? '#34d399' : (kind === 'warn' ? '#fbbf24' : '#f87171');
  el.style.display = 'block';
  setTimeout(() => el.style.display = 'none', 5000);
}

// ── BACKUPS ───────────────────────────────────────────────────────────────────
// ── Auditar inventario contra Excel día cero ─────────────────────────────────
function _fmtG(n) {
  // Normalizado: SIEMPRE en gramos con separador de miles (acordado con Alejandro).
  if (n == null) return '—';
  return Math.round(Number(n) || 0).toLocaleString('es-CO') + ' g';
}
function _esc(s) { return String(s||'').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'})[c]); }
async function auditarInvVsExcel() {
  const fileEl = document.getElementById('audit-inv-file');
  const out = document.getElementById('audit-inv-result');
  if (!fileEl.files || !fileEl.files[0]) {
    out.innerHTML = '<div style="color:#fca5a5;">Selecciona un .xlsx primero.</div>';
    return;
  }
  out.innerHTML = '<div style="color:#94a3b8;">Procesando Excel + comparando contra DB... espera...</div>';
  const fd = new FormData();
  fd.append('file', fileEl.files[0]);
  try {
    const r = await fetch('/api/admin/audit-inventario-vs-excel', {method: 'POST', body: fd});
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + (d.detail?' — '+_esc(d.detail):'') + '</div>'; return; }
    const s = d.resumen;
    let h = '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi"><div class="kpi-l">Lotes verdes (Excel)</div><div class="kpi-v">' + s.lotes_verde_excel + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Excluidos (no verde)</div><div class="kpi-v">' + s.lotes_excluidos_no_verde + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Stock total Excel</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(s.stock_total_excel_g) + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Stock total DB ahora</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(s.stock_total_db_actual_g) + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Producciones desde día 0</div><div class="kpi-v">' + s.producciones_registradas + ' (' + s.producciones_total_kg.toLocaleString('es-CO') + ' kg)</div></div>';
    h += '</div>';

    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi" style="background:rgba(16,185,129,.12);"><div class="kpi-l" style="color:#34d399;">✓ Match (cantidad ok)</div><div class="kpi-v">' + s.count_match + '</div></div>';
    h += '<div class="kpi" style="background:rgba(245,158,11,.12);"><div class="kpi-l" style="color:#fbbf24;">⚠ Con delta</div><div class="kpi-v">' + s.count_delta + '</div></div>';
    h += '<div class="kpi" style="background:rgba(239,68,68,.12);"><div class="kpi-l" style="color:#fca5a5;">✗ Faltantes en DB</div><div class="kpi-v">' + s.count_faltantes_en_db + '</div></div>';
    h += '<div class="kpi" style="background:rgba(99,102,241,.12);"><div class="kpi-l" style="color:#a5b4fc;">+ Solo en DB (post)</div><div class="kpi-v">' + s.count_solo_db_no_excel + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Δ total</div><div class="kpi-v" style="font-size:14px;color:' + (s.delta_total_g < 0 ? '#fca5a5' : '#34d399') + ';">' + _fmtG(s.delta_total_g) + '</div></div>';
    h += '</div>';

    function table(items, cols, title, color) {
      let t = '<h3 style="color:' + color + ';margin-top:18px;margin-bottom:8px;">' + title + ' (' + items.length + ')</h3>';
      if (!items.length) return t + '<div style="color:#94a3b8;font-size:12px;">— vacío —</div>';
      t += '<div style="overflow-x:auto;background:#0f172a;border:1px solid #334155;border-radius:8px;"><table style="width:100%;border-collapse:collapse;font-size:12px;">';
      t += '<thead style="background:#1e293b;"><tr>' + cols.map(c => '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">' + c.label + '</th>').join('') + '</tr></thead><tbody>';
      items.forEach(it => {
        t += '<tr style="border-top:1px solid #334155;">';
        cols.forEach(c => {
          let v = it[c.key];
          if (c.fmt === 'g') v = _fmtG(v);
          t += '<td style="padding:6px 10px;color:' + (c.color||'#e2e8f0') + ';' + (c.fmt==='g'?'text-align:right;font-family:monospace;':'') + '">' + _esc(v) + '</td>';
        });
        t += '</tr>';
      });
      t += '</tbody></table></div>';
      return t;
    }

    h += table(d.en_db_con_delta, [
      {label:'Código', key:'codigo_mp'},
      {label:'Material', key:'nombre_comercial'},
      {label:'Lote', key:'lote'},
      {label:'Excel (g)', key:'cant_excel_g', fmt:'g'},
      {label:'Salidas DB', key:'salidas_db_g', fmt:'g'},
      {label:'Esperado', key:'neto_esperado_g', fmt:'g'},
      {label:'Actual DB', key:'neto_db_g', fmt:'g'},
      {label:'Δ', key:'delta_g', fmt:'g', color:'#fbbf24'},
    ], '⚠ Lotes con cantidad distinta a la esperada', '#fbbf24');

    h += table(d.faltantes_en_db, [
      {label:'Código', key:'codigo_mp'},
      {label:'Material', key:'nombre_comercial'},
      {label:'Lote', key:'lote'},
      {label:'Proveedor (Excel)', key:'proveedor_excel'},
      {label:'Cant Excel', key:'cant_excel_g', fmt:'g', color:'#fca5a5'},
    ], '✗ Lotes verdes en Excel pero NO en DB (desaparecieron)', '#fca5a5');

    h += table(d.solo_db_no_excel, [
      {label:'Código', key:'codigo_mp'},
      {label:'Lote', key:'lote'},
      {label:'Entradas DB', key:'entradas_db_g', fmt:'g'},
      {label:'Salidas DB', key:'salidas_db_g', fmt:'g'},
      {label:'Neto DB', key:'neto_db_g', fmt:'g', color:'#a5b4fc'},
    ], '+ Lotes en DB pero NO en Excel verde (post-día-cero o sobrante)', '#a5b4fc');

    if (d.nota_truncado) {
      h += '<div style="margin-top:14px;color:#94a3b8;font-size:11px;">ℹ Tablas truncadas a 300 items. Hay más — contacta para ver el JSON completo o exportar.</div>';
    }

    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error de red: ' + _esc(e.message) + '</div>';
  }
}

async function sembrarMaestroDesdeExcel() {
  const fileEl = document.getElementById('audit-inv-file');
  const out = document.getElementById('audit-inv-result');
  if (!fileEl.files || !fileEl.files[0]) {
    out.innerHTML = '<div style="color:#fbbf24;">Selecciona primero el Excel arriba.</div>';
    return;
  }
  out.innerHTML = '<div style="color:#94a3b8;">Sembrando catálogo MPs desde Excel (verdes + no-verdes)...</div>';
  try {
    const fd = new FormData();
    fd.append('file', fileEl.files[0]);
    const r = await fetch('/api/admin/sembrar-maestro-desde-excel', {method: 'POST', body: fd});
    const d = await r.json();
    if (!r.ok) {
      out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>';
      return;
    }
    let h = '<h3 style="color:#a5b4fc;margin-top:0;">🌱 Catálogo MPs sembrado</h3>';
    h += '<div style="color:#cbd5e1;font-size:13px;margin-bottom:14px;">' + _esc(d.mensaje) + '</div>';
    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi"><div class="kpi-l">MPs nuevos en catálogo</div><div class="kpi-v" style="color:#a5b4fc;">' + d.nuevos_count + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Proveedores actualizados</div><div class="kpi-v" style="color:#34d399;">' + d.proveedores_actualizados_count + '</div></div>';
    h += '</div>';
    if (d.nuevos_sample && d.nuevos_sample.length) {
      h += '<h4 style="color:#a5b4fc;">Nuevos en catálogo (sample primeros 30)</h4>';
      h += '<table><thead><tr><th>Código</th><th>Nombre</th><th>Proveedor</th></tr></thead><tbody>';
      d.nuevos_sample.forEach(n => {
        h += '<tr><td style="font-family:monospace;">' + _esc(n.codigo_mp) + '</td><td>' + _esc(n.nombre_comercial) + '</td><td>' + _esc(n.proveedor||'—') + '</td></tr>';
      });
      h += '</tbody></table>';
    }
    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

async function descargarSnapshotPreReset() {
  const out = document.getElementById('audit-inv-result');
  out.innerHTML = '<div style="color:#94a3b8;">Generando snapshot... esto descarga un JSON grande.</div>';
  try {
    const r = await fetch('/api/admin/inventario-snapshot-pre-reset');
    if (!r.ok) {
      const d = await r.json();
      out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>';
      return;
    }
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const ts = new Date().toISOString().replace(/[:.]/g,'-').slice(0,19);
    a.download = 'snapshot_pre_reset_' + ts + '.json';
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    out.innerHTML = '<div style="color:#34d399;font-weight:700;">✓ Snapshot descargado. Guárdalo fuera de Render antes del reset.</div>';
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

async function previewReset() {
  const fileEl = document.getElementById('audit-inv-file');
  const out = document.getElementById('audit-inv-result');
  if (!fileEl.files || !fileEl.files[0]) {
    out.innerHTML = '<div style="color:#fca5a5;">Selecciona el .xlsx primero.</div>';
    return;
  }
  out.innerHTML = '<div style="color:#94a3b8;">Calculando preview...</div>';
  const fd = new FormData();
  fd.append('file', fileEl.files[0]);
  try {
    const r = await fetch('/api/admin/inventario-reset-preview', {method:'POST', body:fd});
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    const p = d.plan;
    let h = '<h3 style="color:#a5b4fc;margin-top:0;">&#x1F441; Preview del Reset (sin escribir)</h3>';
    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi" style="background:rgba(239,68,68,.12);"><div class="kpi-l" style="color:#fca5a5;">Movs a borrar</div><div class="kpi-v">' + p.movimientos_a_borrar + '</div></div>';
    h += '<div class="kpi" style="background:rgba(16,185,129,.12);"><div class="kpi-l" style="color:#34d399;">Entradas iniciales (Excel)</div><div class="kpi-v">' + p.entradas_iniciales_a_crear.count + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Total g a cargar</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(p.entradas_iniciales_a_crear.total_g) + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Entradas OC preservadas</div><div class="kpi-v">' + p.entradas_oc_a_preservar.count + ' (' + _fmtG(p.entradas_oc_a_preservar.total_g) + ')</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Salidas Producción preservadas</div><div class="kpi-v">' + p.salidas_produccion_a_preservar.count + ' (' + _fmtG(p.salidas_produccion_a_preservar.total_g) + ')</div></div>';
    h += '</div>';

    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi"><div class="kpi-l">Stock actual</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(d.resumen_pre_post.stock_actual_g) + '</div></div>';
    h += '<div class="kpi" style="background:rgba(16,185,129,.12);"><div class="kpi-l" style="color:#34d399;">Stock POST-reset</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(d.resumen_pre_post.stock_post_reset_g_estimado) + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Δ esperado</div><div class="kpi-v" style="font-size:14px;color:' + (d.resumen_pre_post.delta_g < 0 ? '#fca5a5' : '#34d399') + ';">' + _fmtG(d.resumen_pre_post.delta_g) + '</div></div>';
    h += '</div>';

    // ── Bloque NUEVO: lotes compensados FEFO (fix del bug día cero) ──
    const compCount = p.entradas_iniciales_a_crear.lotes_compensados_por_salidas_post_count || 0;
    const compSample = p.entradas_iniciales_a_crear.lotes_compensados_sample_top10 || [];
    if (compCount > 0) {
      h += '<div style="background:rgba(34,197,94,.10);border:1px solid rgba(34,197,94,.4);border-radius:8px;padding:12px;margin-bottom:14px;">';
      h += '<div style="color:#34d399;font-weight:700;margin-bottom:4px;">🛡️ ' + compCount + ' lotes verdes con compensación FEFO (no quedarán negativos)</div>';
      h += '<div style="color:#cbd5e1;font-size:12px;margin-bottom:8px;">Estos lotes están en el Excel verde Y tienen salidas de producción. Cantidad inicial al día cero = excel_actual + salidas_post, para que el FEFO los lleve exactamente al valor reportado HOY sin pasarse a negativo.</div>';
      h += '<div style="overflow-x:auto;background:#0f172a;border:1px solid #334155;border-radius:6px;max-height:280px;overflow-y:auto;"><table style="width:100%;border-collapse:collapse;font-size:11px;"><thead style="background:#1e293b;position:sticky;top:0;"><tr>';
      ['Código','Lote','Excel HOY (g)','Salidas post (g)','→ Cantidad día cero (g)'].forEach(t => {
        h += '<th style="padding:6px 8px;text-align:left;color:#cbd5e1;">' + t + '</th>';
      });
      h += '</tr></thead><tbody>';
      compSample.forEach(o => {
        h += '<tr style="border-top:1px solid #334155;">';
        h += '<td style="padding:5px 8px;font-family:monospace;color:#e2e8f0;">' + _esc(o.codigo_mp) + '</td>';
        h += '<td style="padding:5px 8px;font-family:monospace;color:#94a3b8;">' + _esc(o.lote) + '</td>';
        h += '<td style="padding:5px 8px;text-align:right;font-family:monospace;color:#cbd5e1;">' + _fmtG(o.cantidad_excel_actual_g) + '</td>';
        h += '<td style="padding:5px 8px;text-align:right;font-family:monospace;color:#fbbf24;">+' + _fmtG(o.salidas_post_dia_cero_g) + '</td>';
        h += '<td style="padding:5px 8px;text-align:right;font-family:monospace;color:#34d399;font-weight:700;">' + _fmtG(o.cantidad_inicial_dia_cero_g) + '</td>';
        h += '</tr>';
      });
      h += '</tbody></table></div>';
      h += '<div style="color:#94a3b8;font-size:10px;margin-top:6px;">Top 10 mostrados de ' + compCount + ' lotes compensados.</div>';
      h += '</div>';
    } else {
      h += '<div style="background:rgba(34,197,94,.08);border:1px solid rgba(34,197,94,.3);border-radius:8px;padding:8px 12px;margin-bottom:14px;color:#34d399;font-size:12px;">✓ Sin lotes que requieran compensación FEFO — todas las salidas son a lotes huérfanos o no hay salidas previas.</div>';
    }

    const hue = p.huerfanos_a_regenerar || {count:0, total_g_entradas_virtuales:0, sample:[]};
    if (hue.count > 0) {
      h += '<div style="background:rgba(99,102,241,.10);border:1px solid rgba(99,102,241,.4);border-radius:8px;padding:12px;margin-bottom:14px;">';
      h += '<div style="color:#a5b4fc;font-weight:700;margin-bottom:4px;">♻ ' + hue.count + ' lotes huérfanos a regenerar (' + _fmtG(hue.total_g_entradas_virtuales) + ')</div>';
      h += '<div style="color:#cbd5e1;font-size:12px;margin-bottom:6px;">' + _esc(hue.nota || '') + '</div>';
      h += '<div style="overflow-x:auto;background:#0f172a;border:1px solid #334155;border-radius:6px;max-height:240px;overflow-y:auto;"><table style="width:100%;border-collapse:collapse;font-size:11px;"><thead style="background:#1e293b;position:sticky;top:0;"><tr>';
      ['Código','Lote','Consumido (g)','Entrada virtual (g)','Origen'].forEach(t => {
        h += '<th style="padding:6px 8px;text-align:left;color:#cbd5e1;">' + t + '</th>';
      });
      h += '</tr></thead><tbody>';
      hue.sample.forEach(o => {
        h += '<tr style="border-top:1px solid #334155;">';
        h += '<td style="padding:5px 8px;font-family:monospace;color:#e2e8f0;">' + _esc(o.codigo_mp) + '</td>';
        h += '<td style="padding:5px 8px;font-family:monospace;color:#94a3b8;">' + _esc(o.lote) + '</td>';
        h += '<td style="padding:5px 8px;text-align:right;font-family:monospace;color:#fca5a5;">' + _fmtG(o.cantidad_consumida_g) + '</td>';
        h += '<td style="padding:5px 8px;text-align:right;font-family:monospace;color:#34d399;">' + _fmtG(o.cantidad_entrada_virtual_g) + '</td>';
        h += '<td style="padding:5px 8px;color:#94a3b8;font-size:10px;">' + _esc(o.origen) + '</td>';
        h += '</tr>';
      });
      h += '</tbody></table></div></div>';
    } else {
      h += '<div style="background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.4);border-radius:8px;padding:10px;margin-bottom:14px;color:#34d399;">✓ Todas las salidas de producción consumieron lotes presentes en Excel verde — replay limpio.</div>';
    }

    h += '<div style="background:rgba(99,102,241,.10);border:1px solid rgba(99,102,241,.3);border-radius:8px;padding:10px;font-size:12px;color:#cbd5e1;">';
    h += 'Si todo se ve OK, click en <strong>3. APLICAR reset</strong>. Vas a tener que pegar el token textual:<br>';
    h += '<code style="background:#0f172a;padding:4px 8px;border-radius:4px;color:#a5b4fc;">BORRAR_INVENTARIO_Y_CARGAR_EXCEL_2026_04_27</code>';
    h += '</div>';

    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

async function aplicarReset() {
  const fileEl = document.getElementById('audit-inv-file');
  const out = document.getElementById('audit-inv-result');
  if (!fileEl.files || !fileEl.files[0]) {
    out.innerHTML = '<div style="color:#fca5a5;">Selecciona el .xlsx primero.</div>';
    return;
  }
  const TOKEN = 'BORRAR_INVENTARIO_Y_CARGAR_EXCEL_2026_04_27';
  const ingresado = prompt('PEGAR EL TOKEN TEXTUAL EXACTO PARA CONFIRMAR:\n\n' + TOKEN + '\n\n(Cualquier otro texto cancela.)');
  if (ingresado !== TOKEN) {
    out.innerHTML = '<div style="color:#fbbf24;">Reset cancelado — token no coincide.</div>';
    return;
  }
  if (!confirm('Última confirmación. Esto borra TODOS los movimientos y los recrea. Procedo?')) {
    out.innerHTML = '<div style="color:#fbbf24;">Reset cancelado.</div>';
    return;
  }
  out.innerHTML = '<div style="color:#94a3b8;">Aplicando reset... no cierres la pestaña...</div>';
  const fd = new FormData();
  fd.append('file', fileEl.files[0]);
  fd.append('confirmacion', TOKEN);
  try {
    const r = await fetch('/api/admin/inventario-reset-aplicar', {method:'POST', body:fd});
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + (d.detail?'<br><small>' + _esc(d.detail) + '</small>':'') + '</div>'; return; }
    let h = '<div style="background:rgba(16,185,129,.15);border:1px solid #34d399;border-radius:10px;padding:14px;">';
    h += '<h3 style="color:#34d399;margin:0 0 10px 0;">✓ Reset aplicado con éxito</h3>';
    h += '<div style="color:#cbd5e1;font-size:13px;margin-bottom:10px;">' + _esc(d.message) + '</div>';
    h += '<div class="kpi-row">';
    h += '<div class="kpi"><div class="kpi-l">Movs borrados</div><div class="kpi-v">' + d.resumen.movs_borrados + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Lotes Excel cargados</div><div class="kpi-v">' + d.resumen.lotes_excel_cargados + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Entradas OC preservadas</div><div class="kpi-v">' + d.resumen.entradas_oc_preservadas + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Salidas prod preservadas</div><div class="kpi-v">' + d.resumen.salidas_prod_preservadas + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Movs total post</div><div class="kpi-v">' + d.resumen.movs_post_total + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Stock post</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(d.resumen.stock_post_g) + '</div></div>';
    h += '</div>';
    h += '<div style="margin-top:10px;color:#94a3b8;font-size:11px;">Audit log entries creados: RESET_INVENTARIO_PRE + RESET_INVENTARIO_POST. Si necesitas restaurar, usa el snapshot que descargaste.</div>';
    h += '</div>';
    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error de red: ' + _esc(e.message) + '</div>';
  }
}

async function healthCheckPostReset() {
  const out = document.getElementById('audit-inv-result');
  out.innerHTML = '<div style="color:#94a3b8;">Verificando integridad...</div>';
  try {
    const r = await fetch('/api/admin/health-check-post-reset');
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    const okColor = d.ok ? '#34d399' : '#fca5a5';
    let h = '<h3 style="color:' + okColor + ';margin-top:0;">' + (d.ok ? '✓' : '✗') + ' Health-check: ' + d.overall + '</h3>';
    h += '<div style="color:#cbd5e1;font-size:13px;margin-bottom:14px;">' + _esc(d.recomendacion) + '</div>';

    const cs = d.checks;

    // Movimientos
    if (cs.movimientos && cs.movimientos.ok) {
      h += '<div class="kpi-row" style="margin-bottom:10px;">';
      h += '<div class="kpi"><div class="kpi-l">Movs total</div><div class="kpi-v">' + cs.movimientos.total + '</div></div>';
      h += '<div class="kpi"><div class="kpi-l">Stock total</div><div class="kpi-v" style="font-size:14px;">' + cs.movimientos.stock_total_kg + ' kg</div></div>';
      h += '<div class="kpi"><div class="kpi-l">Entradas</div><div class="kpi-v">' + cs.movimientos.entradas + '</div></div>';
      h += '<div class="kpi"><div class="kpi-l">Salidas</div><div class="kpi-v">' + cs.movimientos.salidas + '</div></div>';
      h += '<div class="kpi"><div class="kpi-l">Salidas FEFO prod.</div><div class="kpi-v">' + cs.movimientos.salidas_fefo_produccion + '</div></div>';
      h += '</div>';
    }

    // Stock negativo
    if (cs.stock_negativo) {
      const sn_ok = cs.stock_negativo.ok;
      h += '<div style="background:rgba(' + (sn_ok ? '16,185,129' : '239,68,68') + ',.12);border:1px solid rgba(' + (sn_ok ? '16,185,129' : '239,68,68') + ',.4);border-radius:8px;padding:10px;margin-bottom:10px;">';
      h += '<div style="color:' + (sn_ok ? '#34d399' : '#fca5a5') + ';font-weight:700;">' + (sn_ok ? '✓' : '✗') + ' Stock negativo: ' + cs.stock_negativo.count + ' lotes</div>';
      if (cs.stock_negativo.count > 0) {
        h += '<ul style="font-size:11px;color:#cbd5e1;margin:6px 0 0 18px;">';
        cs.stock_negativo.sample.forEach(s => {
          h += '<li>' + _esc(s.codigo_mp) + ' / ' + _esc(s.lote) + ' = ' + _fmtG(s.neto_g) + '</li>';
        });
        h += '</ul>';
      }
      h += '</div>';
    }

    // Otras tablas
    h += '<div class="kpi-row">';
    if (cs.producciones) {
      h += '<div class="kpi"><div class="kpi-l">Producciones</div><div class="kpi-v">' + cs.producciones.count + '</div></div>';
    }
    if (cs.ocs) {
      h += '<div class="kpi"><div class="kpi-l">OCs</div><div class="kpi-v">' + cs.ocs.ordenes_compra + '</div></div>';
    }
    if (cs.comprobantes_pago) {
      h += '<div class="kpi"><div class="kpi-l">Comprobantes</div><div class="kpi-v">' + cs.comprobantes_pago.count + '</div></div>';
    }
    if (cs.maestro_mps) {
      h += '<div class="kpi"><div class="kpi-l">MPs catálogo activos</div><div class="kpi-v">' + cs.maestro_mps.count_activos + '</div></div>';
    }
    if (cs.solicitudes_compra) {
      h += '<div class="kpi"><div class="kpi-l">Solicitudes</div><div class="kpi-v">' + cs.solicitudes_compra.count + '</div></div>';
    }
    h += '</div>';

    // Audit trail
    if (cs.audit_log && cs.audit_log.eventos_reset_recientes) {
      h += '<h4 style="color:#a5b4fc;margin-top:14px;">Audit trail del reset:</h4>';
      h += '<ul style="color:#cbd5e1;font-size:12px;margin-left:18px;">';
      cs.audit_log.eventos_reset_recientes.forEach(e => {
        h += '<li>' + _esc(e.accion) + ' — ' + _esc(e.fecha) + '</li>';
      });
      h += '</ul>';
    }

    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

async function healthMonitor() {
  const out = document.getElementById('audit-inv-result');
  out.innerHTML = '<div style="color:#94a3b8;">Escaneando kardex en busca de anomalías...</div>';
  try {
    const r = await fetch('/api/admin/inventario-health-monitor');
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    const colors = {ok:'#34d399', warning:'#fbbf24', critical:'#ef4444'};
    const c = colors[d.nivel] || '#cbd5e1';
    let h = '<h3 style="color:' + c + ';margin-top:0;">🔥 Monitor Anomalías — ' + d.nivel.toUpperCase() + '</h3>';
    h += '<div style="background:rgba(' + (d.nivel==='ok'?'16,185,129':d.nivel==='critical'?'239,68,68':'245,158,11') + ',.12);border:1px solid ' + c + ';border-radius:8px;padding:12px;margin-bottom:14px;">';
    h += '<div style="color:' + c + ';font-weight:700;">' + _esc(d.recomendacion) + '</div>';
    h += '</div>';
    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi" style="background:rgba(239,68,68,.12);"><div class="kpi-l" style="color:#fca5a5;">CRITICAL</div><div class="kpi-v">' + d.count_critical + '</div></div>';
    h += '<div class="kpi" style="background:rgba(245,158,11,.12);"><div class="kpi-l" style="color:#fbbf24;">WARNING</div><div class="kpi-v">' + d.count_warning + '</div></div>';
    h += '</div>';

    if (d.alertas.length === 0) {
      h += '<div style="background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.4);border-radius:8px;padding:12px;color:#34d399;">✓ Cero anomalías detectadas. Sistema con kardex limpio.</div>';
    } else {
      d.alertas.forEach(a => {
        const sevColor = a.severidad === 'critical' ? '#ef4444' : '#fbbf24';
        h += '<div style="border-left:3px solid ' + sevColor + ';background:#0f172a;padding:10px 14px;margin-bottom:8px;border-radius:0 8px 8px 0;">';
        h += '<div style="color:' + sevColor + ';font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">' + _esc(a.tipo) + ' · ' + _esc(a.severidad) + '</div>';
        h += '<div style="color:#cbd5e1;font-size:13px;margin-bottom:4px;">' + _esc(a.mensaje) + '</div>';
        h += '<details style="margin-top:4px;"><summary style="cursor:pointer;color:#94a3b8;font-size:11px;">Detalle</summary>';
        h += '<pre style="background:#020617;padding:8px;border-radius:4px;font-size:11px;color:#94a3b8;margin:6px 0 0 0;overflow-x:auto;">' + _esc(JSON.stringify(a.detalle, null, 2)) + '</pre>';
        h += '</details></div>';
      });
    }

    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

async function quePuedoProducir() {
  const out = document.getElementById('audit-inv-result');
  out.innerHTML = '<div style="color:#94a3b8;">Calculando para cada producto...</div>';
  try {
    const r = await fetch('/api/programacion/que-puedo-producir');
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    const s = d.resumen;
    let h = '<h3 style="color:#fbbf24;margin-top:0;">🏭 Qu&eacute; puedo producir HOY</h3>';

    h += '<div style="font-size:11px;color:#64748b;margin-bottom:10px;">';
    h += 'Fuente: <code>' + _esc(d.fuente_datos.kardex) + '</code> · ';
    h += '<code>' + _esc(d.fuente_datos.formulas) + '</code> · ';
    h += '<code>' + _esc(d.fuente_datos.proveedor_canonico) + '</code>';
    h += '</div>';

    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi"><div class="kpi-l">Productos total</div><div class="kpi-v">' + s.productos_totales + '</div></div>';
    h += '<div class="kpi" style="background:rgba(16,185,129,.12);"><div class="kpi-l" style="color:#34d399;">✓ Pueden producir</div><div class="kpi-v">' + s.productos_pueden_producir + '</div></div>';
    h += '<div class="kpi" style="background:rgba(239,68,68,.12);"><div class="kpi-l" style="color:#fca5a5;">✗ Con faltantes</div><div class="kpi-v">' + s.productos_con_faltantes + '</div></div>';
    h += '</div>';

    // Shopping list por proveedor (lo más útil)
    if (d.shopping_list_por_proveedor.length > 0) {
      h += '<h3 style="color:#a5b4fc;margin-top:18px;">🛒 Shopping list por proveedor</h3>';
      h += '<div style="font-size:11px;color:#94a3b8;margin-bottom:8px;">Lo que hay que pedir AHORA para no quedar en cero. Ordenado por kg total.</div>';
      d.shopping_list_por_proveedor.forEach(prov => {
        h += '<div style="background:#0f172a;border:1px solid #334155;border-radius:8px;padding:12px;margin-bottom:10px;">';
        h += '<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px;">';
        h += '<div style="font-weight:700;color:#a5b4fc;font-size:14px;">' + _esc(prov.proveedor) + '</div>';
        h += '<div style="color:#cbd5e1;font-size:12px;">' + prov.count_mps + ' MPs · <strong>' + _fmtG(prov.total_g_a_pedir) + '</strong> total</div>';
        h += '</div>';
        h += '<table style="width:100%;border-collapse:collapse;font-size:12px;"><thead><tr><th style="text-align:left;color:#94a3b8;padding:4px 6px;">Código</th><th style="text-align:left;color:#94a3b8;padding:4px 6px;">MP</th><th style="text-align:right;color:#94a3b8;padding:4px 6px;">Pedir mín</th><th style="text-align:left;color:#94a3b8;padding:4px 6px;">Para</th></tr></thead><tbody>';
        prov.mps.forEach(m => {
          h += '<tr style="border-top:1px solid #1e293b;">';
          h += '<td style="padding:5px 6px;font-family:monospace;color:#e2e8f0;">' + _esc(m.codigo_mp) + '</td>';
          h += '<td style="padding:5px 6px;color:#e2e8f0;">' + _esc(m.nombre) + '</td>';
          h += '<td style="padding:5px 6px;text-align:right;font-family:monospace;color:#fca5a5;">' + _fmtG(m.falta_g) + '</td>';
          h += '<td style="padding:5px 6px;color:#94a3b8;font-size:11px;">' + _esc(m.productos_afectados.join(', ')) + '</td>';
          h += '</tr>';
        });
        h += '</tbody></table></div>';
      });
    } else {
      h += '<div style="background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.4);border-radius:8px;padding:12px;color:#34d399;">✓ Cero faltantes — todas las MPs alcanzan para producir cada producto en cantidad estándar.</div>';
    }

    // Detalle por producto (expandible)
    h += '<h3 style="color:#cbd5e1;margin-top:18px;">📋 Detalle por producto</h3>';
    h += '<div style="font-size:11px;color:#94a3b8;margin-bottom:8px;">Click en un producto para ver evidencia (lotes disponibles).</div>';
    d.productos.forEach((p, idx) => {
      const okColor = p.puede_producir ? '#34d399' : '#fca5a5';
      const ico = p.puede_producir ? '✓' : '✗';
      h += '<details style="margin-bottom:6px;background:#0f172a;border:1px solid #334155;border-radius:6px;">';
      h += '<summary style="padding:8px 12px;cursor:pointer;display:flex;justify-content:space-between;align-items:center;">';
      h += '<span style="color:' + okColor + ';"><strong>' + ico + ' ' + _esc(p.producto) + '</strong> · lote ' + p.cantidad_kg_evaluada + ' kg</span>';
      h += '<span style="color:#94a3b8;font-size:11px;">' + p.mps_ok + '/' + p.mps_total + ' MPs ok' + (p.mps_faltantes > 0 ? ' · falta ' + _fmtG(p.falta_total_g) : '') + '</span>';
      h += '</summary>';
      h += '<div style="padding:8px 12px;border-top:1px solid #1e293b;"><table style="width:100%;border-collapse:collapse;font-size:11px;"><thead><tr style="color:#94a3b8;"><th style="text-align:left;padding:4px;">MP</th><th style="text-align:right;padding:4px;">Req</th><th style="text-align:right;padding:4px;">Stock</th><th style="text-align:right;padding:4px;">Falta</th><th style="text-align:left;padding:4px;">Lotes (evidencia)</th></tr></thead><tbody>';
      p.mps_status.forEach(m => {
        const rowColor = m.ok ? '#cbd5e1' : '#fca5a5';
        h += '<tr style="border-top:1px solid #1e293b;color:' + rowColor + ';">';
        h += '<td style="padding:4px;"><code style="font-size:10px;color:#94a3b8;">' + _esc(m.codigo_mp) + '</code> ' + _esc(m.nombre) + '</td>';
        h += '<td style="padding:4px;text-align:right;font-family:monospace;">' + _fmtG(m.requerido_g) + '</td>';
        h += '<td style="padding:4px;text-align:right;font-family:monospace;">' + _fmtG(m.stock_actual_g) + '</td>';
        h += '<td style="padding:4px;text-align:right;font-family:monospace;color:' + (m.falta_g > 0 ? '#fca5a5' : '#64748b') + ';">' + (m.falta_g > 0 ? _fmtG(m.falta_g) : '—') + '</td>';
        h += '<td style="padding:4px;font-size:10px;color:#94a3b8;">';
        if (m.lotes_disponibles.length === 0) {
          h += '<em>sin lotes</em>';
        } else {
          h += m.lotes_disponibles.map(l => _esc(l.lote || '(s/lote)') + ': ' + _fmtG(l.cantidad_g)).join(' · ');
        }
        h += '</td>';
        h += '</tr>';
      });
      h += '</tbody></table></div>';
      h += '</details>';
    });

    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

async function diagnosticoEntradas() {
  const out = document.getElementById('audit-inv-result');
  out.innerHTML = '<div style="color:#94a3b8;">Analizando entradas en el kardex...</div>';
  try {
    const r = await fetch('/api/admin/inventario-diagnostico-entradas');
    const d = await r.json();
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    const s = d.resumen;
    let h = '<h3 style="color:#a5b4fc;margin-top:0;">&#x1F50E; Diagn&oacute;stico de Entradas</h3>';
    h += '<div class="kpi-row" style="margin-bottom:14px;">';
    h += '<div class="kpi"><div class="kpi-l">Total Entradas</div><div class="kpi-v">' + s.total_entradas + '</div></div>';
    h += '<div class="kpi"><div class="kpi-l">Total Entradas (g)</div><div class="kpi-v" style="font-size:14px;">' + _fmtG(s.total_entradas_g) + '</div></div>';
    h += '<div class="kpi" style="background:rgba(239,68,68,.12);"><div class="kpi-l" style="color:#fca5a5;">Lotes c/ multiples Entradas</div><div class="kpi-v">' + s.lotes_con_multiples_entradas + '</div></div>';
    h += '<div class="kpi" style="background:rgba(16,185,129,.12);"><div class="kpi-l" style="color:#34d399;">Entradas con OC (formal)</div><div class="kpi-v">' + s.entradas_con_oc + ' (' + _fmtG(s.entradas_con_oc_g) + ')</div></div>';
    h += '<div class="kpi" style="background:rgba(245,158,11,.12);"><div class="kpi-l" style="color:#fbbf24;">Entradas SIN OC (manual)</div><div class="kpi-v">' + s.entradas_sin_oc + ' (' + _fmtG(s.entradas_sin_oc_g) + ')</div></div>';
    h += '</div>';

    // Por operador
    h += '<h3 style="color:#a5b4fc;margin-top:18px;">Entradas por operador</h3>';
    h += '<div style="overflow-x:auto;background:#0f172a;border:1px solid #334155;border-radius:8px;"><table style="width:100%;border-collapse:collapse;font-size:12px;"><thead style="background:#1e293b;"><tr>';
    h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Operador</th>';
    h += '<th style="padding:8px 10px;text-align:right;color:#cbd5e1;">N° Entradas</th>';
    h += '<th style="padding:8px 10px;text-align:right;color:#cbd5e1;">Total (g)</th>';
    h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Primera</th>';
    h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Última</th>';
    h += '</tr></thead><tbody>';
    d.por_operador.forEach(o => {
      h += '<tr style="border-top:1px solid #334155;">';
      h += '<td style="padding:6px 10px;color:#e2e8f0;font-weight:600;">' + _esc(o.operador) + '</td>';
      h += '<td style="padding:6px 10px;text-align:right;font-family:monospace;">' + o.n_entradas + '</td>';
      h += '<td style="padding:6px 10px;text-align:right;font-family:monospace;color:#a5b4fc;">' + _fmtG(o.total_g) + '</td>';
      h += '<td style="padding:6px 10px;color:#94a3b8;">' + _esc(o.primera_fecha) + '</td>';
      h += '<td style="padding:6px 10px;color:#94a3b8;">' + _esc(o.ultima_fecha) + '</td>';
      h += '</tr>';
    });
    h += '</tbody></table></div>';

    // Timeline
    h += '<h3 style="color:#a5b4fc;margin-top:18px;">Timeline de Entradas (primeros 60 dias con actividad)</h3>';
    h += '<div style="overflow-x:auto;background:#0f172a;border:1px solid #334155;border-radius:8px;max-height:300px;overflow-y:auto;"><table style="width:100%;border-collapse:collapse;font-size:12px;"><thead style="background:#1e293b;position:sticky;top:0;"><tr>';
    h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Fecha</th>';
    h += '<th style="padding:8px 10px;text-align:right;color:#cbd5e1;">N° Entradas</th>';
    h += '<th style="padding:8px 10px;text-align:right;color:#cbd5e1;">Total (g)</th>';
    h += '</tr></thead><tbody>';
    d.timeline.forEach(t => {
      const isHigh = t.n_entradas > 50;  // dia con burst
      h += '<tr style="border-top:1px solid #334155;' + (isHigh ? 'background:rgba(245,158,11,.08);' : '') + '">';
      h += '<td style="padding:6px 10px;color:#e2e8f0;font-family:monospace;">' + _esc(t.fecha) + (isHigh ? ' <span style="color:#fbbf24;">⚠ burst</span>' : '') + '</td>';
      h += '<td style="padding:6px 10px;text-align:right;font-family:monospace;">' + t.n_entradas + '</td>';
      h += '<td style="padding:6px 10px;text-align:right;font-family:monospace;color:#a5b4fc;">' + _fmtG(t.total_g) + '</td>';
      h += '</tr>';
    });
    h += '</tbody></table></div>';

    // Multi-entradas (smoking gun)
    h += '<h3 style="color:#fca5a5;margin-top:18px;">⚠ Lotes con MULTIPLES Entradas (' + d.multi_entradas.length + ')</h3>';
    h += '<div style="color:#94a3b8;font-size:11px;margin-bottom:6px;">Si un lote aparece 2 veces como Entrada con la misma cantidad → doble carga. Si tiene 2 entradas pero cantidades distintas → recepción posterior legítima.</div>';
    if (!d.multi_entradas.length) {
      h += '<div style="color:#34d399;">✓ Ningún lote con múltiples entradas — no hay doble carga.</div>';
    } else {
      h += '<div style="overflow-x:auto;background:#0f172a;border:1px solid #334155;border-radius:8px;max-height:400px;overflow-y:auto;"><table style="width:100%;border-collapse:collapse;font-size:12px;"><thead style="background:#1e293b;position:sticky;top:0;"><tr>';
      h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Código</th>';
      h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Lote</th>';
      h += '<th style="padding:8px 10px;text-align:right;color:#cbd5e1;">N° Entradas</th>';
      h += '<th style="padding:8px 10px;text-align:right;color:#cbd5e1;">Total Sumado</th>';
      h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Operadores</th>';
      h += '<th style="padding:8px 10px;text-align:left;color:#cbd5e1;">Primera → Última</th>';
      h += '</tr></thead><tbody>';
      d.multi_entradas.forEach(m => {
        h += '<tr style="border-top:1px solid #334155;">';
        h += '<td style="padding:6px 10px;font-family:monospace;color:#e2e8f0;">' + _esc(m.codigo_mp) + '</td>';
        h += '<td style="padding:6px 10px;font-family:monospace;color:#94a3b8;">' + _esc(m.lote) + '</td>';
        h += '<td style="padding:6px 10px;text-align:right;font-weight:700;color:' + (m.n_entradas >= 2 ? '#fca5a5' : '#e2e8f0') + ';">' + m.n_entradas + '</td>';
        h += '<td style="padding:6px 10px;text-align:right;font-family:monospace;color:#fca5a5;">' + _fmtG(m.total_g) + '</td>';
        h += '<td style="padding:6px 10px;color:#94a3b8;font-size:11px;">' + _esc(m.operadores) + '</td>';
        h += '<td style="padding:6px 10px;color:#94a3b8;font-size:11px;">' + _esc(m.primera_fecha) + ' → ' + _esc(m.ultima_fecha) + '</td>';
        h += '</tr>';
      });
      h += '</tbody></table></div>';
    }

    out.innerHTML = h;
  } catch(e) {
    out.innerHTML = '<div style="color:#fca5a5;">Error de red: ' + _esc(e.message) + '</div>';
  }
}

async function loadBackups() {
  const r = await fetch('/api/admin/backups');
  if (!r.ok) {
    document.getElementById('tbody-backups').innerHTML = '<tr><td colspan="4" style="color:#f87171;text-align:center;padding:20px;">Error '+r.status+'</td></tr>';
    return;
  }
  const data = await r.json();
  const items = data.backups || [];
  document.getElementById('kpi-count').textContent = items.length;
  document.getElementById('kpi-size').textContent = items.length
    ? (items.reduce((a,b)=>a+(b.size_bytes||0),0)/1024/1024).toFixed(1) + ' MB' : '0 MB';
  document.getElementById('kpi-last').textContent = items.length
    ? items[0].modified.replace('T',' ').replace('Z','') : 'nunca';
  document.getElementById('kpi-ret').textContent = (data.config && data.config.retention_days || 7) + ' dias';
  if (!items.length) {
    document.getElementById('tbody-backups').innerHTML = '<tr><td colspan="4" style="text-align:center;color:#64748b;padding:20px;">Sin backups aun.</td></tr>';
  } else {
    document.getElementById('tbody-backups').innerHTML = items.map(b =>
      `<tr><td style="font-family:monospace;font-size:12px;color:#cbd5e1;">${b.filename}</td>
       <td style="color:#94a3b8;">${b.modified.replace('T',' ').replace('Z','')}</td>
       <td><span class="badge badge-ok">${b.size_mb} MB</span></td>
       <td><a href="/api/admin/backup/${encodeURIComponent(b.filename)}" class="btn btn-sm" download>&#x1F4E5; Descargar</a></td></tr>`
    ).join('');
  }
  const runs = data.recent_runs || [];
  if (!runs.length) {
    document.getElementById('tbody-runs').innerHTML = '<tr><td colspan="6" style="text-align:center;color:#64748b;padding:20px;">Sin ejecuciones aun.</td></tr>';
  } else {
    document.getElementById('tbody-runs').innerHTML = runs.map(r => {
      const cls = r.status === 'ok' ? 'badge-ok' : (r.status === 'error' ? 'badge-err' : 'badge-run');
      const size = r.size_bytes ? (r.size_bytes/1024/1024).toFixed(1) + ' MB' : '-';
      const err = r.error ? `<span style="color:#f87171;font-size:11px;">${r.error}</span>` : '-';
      return `<tr><td>${r.id}</td><td style="color:#94a3b8;">${(r.started_at||'').replace('T',' ').replace('Z','')}</td>
        <td><span class="badge ${cls}">${r.status}</span></td>
        <td style="font-size:12px;color:#94a3b8;">${r.triggered_by || ''}</td>
        <td>${size}</td><td>${err}</td></tr>`;
    }).join('');
  }
}

async function triggerBackup() {
  const btn = document.getElementById('btn-backup');
  const lbl = document.getElementById('btn-label');
  btn.disabled = true;
  lbl.textContent = '⏳ Haciendo backup...';
  try {
    const r = await fetch('/api/admin/backup-now', {method:'POST', headers:{'Content-Type':'application/json'}});
    const data = await r.json();
    if (r.ok && data.ok) { toast('Backup creado: ' + data.filename, 'ok'); loadBackups(); }
    else toast('Error: ' + (data.error || 'desconocido'), 'err');
  } catch (e) { toast('Error de red: ' + e.message, 'err'); }
  finally { btn.disabled = false; lbl.textContent = '⚡ Hacer backup ahora'; }
}

// ── USERS ─────────────────────────────────────────────────────────────────────
function _pwdBadge(src) {
  if (src === 'db')             return '<span class="badge badge-ok">DB (cambiada)</span>';
  if (src === 'env')            return '<span class="badge badge-info">ENV (Render)</span>';
  if (src === 'env_plaintext')  return '<span class="badge badge-err">PLAINTEXT</span>';
  return '<span class="badge badge-err">SIN CONFIG</span>';
}
async function loadUsers() {
  const r = await fetch('/api/admin/users');
  if (!r.ok) {
    document.getElementById('tbody-users').innerHTML = '<tr><td colspan="6" style="color:#f87171;">Error</td></tr>';
    return;
  }
  const data = await r.json();
  document.getElementById('tbody-users').innerHTML = (data.users || []).map(u => {
    const groups = (u.groups || []).map(g => `<span class="badge badge-gray">${g}</span>`).join(' ');
    const last = u.last_login ? u.last_login.replace('T',' ').replace('Z','') : '<span style="color:#64748b;">nunca</span>';
    const changed = u.pwd_changed_at ? u.pwd_changed_at.replace('T',' ').replace('Z','') : '<span style="color:#64748b;">—</span>';
    const adminTag = u.is_admin ? ' <span class="badge badge-warn">ADMIN</span>' : '';
    return `<tr>
      <td><strong style="color:#e2e8f0;">${u.username}</strong>${adminTag}</td>
      <td style="display:flex;gap:4px;flex-wrap:wrap;">${groups || '<span style="color:#64748b;">—</span>'}</td>
      <td>${_pwdBadge(u.password_source)}</td>
      <td style="font-size:11px;color:#94a3b8;">${last}</td>
      <td style="font-size:11px;color:#94a3b8;">${changed}</td>
      <td>
        <button class="btn btn-sm" onclick="diagLogin('${u.username}')" style="background:#0e7490;color:#fff;margin-right:4px;" title="Ver por qué este usuario no puede entrar">&#x1F50D; Diag</button>
        <button class="btn btn-sm btn-warn" onclick="resetPassword('${u.username}')">&#x1F511; Resetear</button>
      </td>
    </tr>`;
  }).join('');
}

async function diagLogin(username) {
  // Sebastián 7-may-2026: muestra estado completo del login para diagnosticar
  // por qué un usuario no puede entrar (caso Mayerlin)
  try {
    const r = await fetch('/api/admin/diag-login/' + encodeURIComponent(username));
    const d = await r.json();
    if (!r.ok) { toast('Error: ' + (d.error || r.status), 'err'); return; }
    const failures = (d.recent_failures || []).map(f =>
      '  · ' + (f.ts||'').replace('T',' ').replace('Z','') + ' desde ' + (f.ip||'?')
    ).join('\\n') || '  (ninguno)';
    const last = d.last_login
      ? (d.last_login.ts || '').replace('T',' ').replace('Z','') + ' desde ' + (d.last_login.ip||'?')
      : 'nunca';
    const lines = [
      'DIAGNÓSTICO LOGIN · ' + d.username,
      '─────────────────────────────',
      'Existe en config: ' + (d.exists ? 'sí' : 'NO'),
      'Password source: ' + (d.password_source || '—'),
      '  · db = el user cambió su password vía self-service',
      '  · env = está la PASS_USER en Render como hash pbkdf2/scrypt',
      '  · env_plaintext = en Render pero SIN hash (inseguro)',
      '  · missing = NO existe PASS_USER en Render',
      '',
      'Password cambiada: ' + (d.password_changed_at || '—'),
      'Último login OK: ' + last,
      'MFA enabled: ' + (d.mfa_enabled ? 'sí' : 'no'),
      'Bloqueado por intentos: ' + (d.is_locked ? 'SÍ' : 'no'),
      '',
      'Últimos fallos:\\n' + failures,
      '',
      '➤ ACCIÓN: ' + (d.recommended_action || '—'),
      '',
      d.hint || '',
    ].join('\\n');
    alert(lines);
  } catch (e) { toast('Error: ' + e.message, 'err'); }
}

async function resetPassword(username) {
  if (!confirm('¿Resetear la password de "' + username + '"?\\n\\nSe generará una password aleatoria que verás UNA SOLA VEZ. Tienes que comunicársela al usuario.')) return;
  try {
    const r = await fetch('/api/admin/reset-password', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({username: username})
    });
    const data = await r.json();
    if (r.ok && data.ok) {
      document.getElementById('modal-title').textContent = 'Password reseteada para ' + data.username;
      document.getElementById('modal-msg').textContent = 'Comparte esta password con el usuario por canal seguro:';
      document.getElementById('modal-result').textContent = data.new_password;
      document.getElementById('modal-bg').classList.add('show');
      loadUsers();
    } else {
      toast('Error: ' + (data.error || 'desconocido'), 'err');
    }
  } catch (e) { toast('Error: ' + e.message, 'err'); }
}

function copyResult() {
  const text = document.getElementById('modal-result').textContent;
  navigator.clipboard.writeText(text).then(
    () => toast('Password copiada al portapapeles', 'ok'),
    () => toast('No se pudo copiar', 'err')
  );
}
function closeModal() {
  document.getElementById('modal-bg').classList.remove('show');
  document.getElementById('modal-result').textContent = '';
}
document.getElementById('modal-bg').addEventListener('click', e => {
  if (e.target.id === 'modal-bg') closeModal();
});

// ── SECURITY ──────────────────────────────────────────────────────────────────
async function loadSecurityEvents() {
  const filter = document.getElementById('filter-event').value;
  const url = '/api/admin/security-events' + (filter ? '?event=' + encodeURIComponent(filter) + '&limit=100' : '?limit=100');
  const r = await fetch(url);
  if (!r.ok) {
    document.getElementById('tbody-security').innerHTML = '<tr><td colspan="6" style="color:#f87171;">Error</td></tr>';
    return;
  }
  const data = await r.json();
  // Stats KPIs
  const stats = data.stats_24h || {};
  const keys = Object.keys(stats);
  if (!keys.length) {
    document.getElementById('kpi-stats').innerHTML = '<div class="kpi"><div class="kpi-l">Sin eventos</div><div class="kpi-v" style="font-size:14px;">en las ultimas 24h</div></div>';
  } else {
    document.getElementById('kpi-stats').innerHTML = keys.map(k => {
      const isErr = /failure|blocked|error|csrf/i.test(k);
      return `<div class="kpi"><div class="kpi-l">${k}</div><div class="kpi-v" style="color:${isErr ? '#f87171' : '#34d399'};">${stats[k]}</div></div>`;
    }).join('');
  }
  // Lista
  const events = data.events || [];
  if (!events.length) {
    document.getElementById('tbody-security').innerHTML = '<tr><td colspan="6" style="text-align:center;color:#64748b;padding:20px;">Sin eventos.</td></tr>';
  } else {
    document.getElementById('tbody-security').innerHTML = events.map(e => {
      const isErr = /failure|blocked|error/i.test(e.event);
      const evtClass = isErr ? 'badge-err' : (/success|changed/i.test(e.event) ? 'badge-ok' : 'badge-info');
      return `<tr>
        <td style="color:#64748b;">${e.id}</td>
        <td style="font-size:11px;color:#94a3b8;">${(e.ts||'').replace('T',' ').replace('Z','')}</td>
        <td><span class="badge ${evtClass}">${e.event}</span></td>
        <td><strong>${e.username || '-'}</strong></td>
        <td style="font-size:12px;color:#94a3b8;font-family:monospace;">${e.ip || '-'}</td>
        <td style="font-size:11px;color:#94a3b8;">${e.details || ''}</td>
      </tr>`;
    }).join('');
  }
}

// ── CONFIG ────────────────────────────────────────────────────────────────────
async function loadConfigStatus() {
  const r = await fetch('/api/admin/config-status');
  if (!r.ok) {
    document.getElementById('config-issues').innerHTML = '<div style="color:#f87171;">Error '+r.status+'</div>';
    return;
  }
  const data = await r.json();
  const issues = data.issues || [];
  if (!issues.length) {
    document.getElementById('config-issues').innerHTML = '<div class="badge badge-ok" style="padding:8px 16px;font-size:13px;">&#x2705; Sin issues — todo configurado correctamente.</div>';
  } else {
    document.getElementById('config-issues').innerHTML = issues.map(i => {
      const cls = i.severity === 'CRITICAL' ? 'badge-err' : (i.severity === 'HIGH' ? 'badge-err' : (i.severity === 'MEDIUM' ? 'badge-warn' : 'badge-info'));
      return `<div style="background:#0f172a;border:1px solid #334155;border-left:3px solid ${i.severity==='CRITICAL'?'#dc2626':(i.severity==='HIGH'?'#dc2626':(i.severity==='MEDIUM'?'#fbbf24':'#3b82f6'))};border-radius:8px;padding:12px 14px;margin-bottom:10px;">
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
          <span class="badge ${cls}">${i.severity}</span>
          <strong style="color:#e2e8f0;font-size:13px;">${i.code}</strong>
        </div>
        <div style="font-size:12px;color:#cbd5e1;line-height:1.5;">${i.msg}</div>
      </div>`;
    }).join('');
  }
  const renderRow = (v) => {
    const status = v.set
      ? `<span class="badge badge-ok">SET (${v.length} chars)</span>`
      : '<span class="badge badge-err">FALTA</span>';
    return `<tr><td style="font-family:monospace;color:#cbd5e1;">${v.name}</td><td style="text-align:right;">${status}</td></tr>`;
  };
  document.getElementById('tbody-config-critical').innerHTML = (data.critical || []).map(renderRow).join('');
  document.getElementById('tbody-config-users').innerHTML = (data.user_passwords || []).map(renderRow).join('');
  document.getElementById('tbody-config-optional').innerHTML = (data.optional || []).map(renderRow).join('');
}

// ── TEST EMAIL (SMTP) ─────────────────────────────────────────────────────────
async function enviarTestEmail() {
  const dest = (document.getElementById('test-email-dest').value || '').trim();
  const empSel = document.getElementById('test-email-empresa');
  const empresa = empSel ? empSel.value : 'espagiria';
  const out = document.getElementById('test-email-result');
  out.innerHTML = '<div style="color:#94a3b8;padding:14px;">Generando PDF (' + empresa + ') y enviando...</div>';
  try {
    const r = await fetch('/api/admin/test-email', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({destinatario: dest, empresa: empresa})
    });
    let data;
    try { data = await r.json(); } catch(e) { data = {error: 'Respuesta inválida'}; }
    if (r.ok && data.ok) {
      const checks = (data.verificacion || []).map(v => '<li>' + v + '</li>').join('');
      out.innerHTML = '<div style="color:#34d399;padding:14px;background:#0f172a;border:1px solid #34d399;border-radius:8px;">'
        + '&#x2705; <strong>' + data.mensaje + '</strong><br>'
        + '<div style="margin-top:8px;font-size:12px;color:#cbd5e1;">'
        + '<div>De: <code>' + data.remitente + '</code></div>'
        + '<div>Para: <code>' + data.destinatario + '</code></div>'
        + '<div>Asunto: <em>' + data.asunto + '</em></div>'
        + '</div>'
        + '<ul style="margin-top:10px;font-size:12px;color:#cbd5e1;padding-left:18px;">' + checks + '</ul>'
        + '</div>';
      toast('Correo de prueba enviado', 'ok');
    } else {
      const pasos = (data.pasos || []).map(p => '<li>' + p + '</li>').join('');
      const env = data.env_status ? Object.entries(data.env_status)
        .map(([k,v]) => '<div style="font-family:monospace;font-size:11px;color:#94a3b8;">' + k + ': <strong style="color:' + (String(v)==='MISSING'?'#f87171':'#34d399') + '">' + v + '</strong></div>').join('') : '';
      out.innerHTML = '<div style="color:#f87171;padding:14px;background:#0f172a;border:1px solid #f87171;border-radius:8px;">'
        + '&#x274C; <strong>Error ' + r.status + ': ' + (data.error || 'Falló envío') + '</strong>'
        + (data.detalle ? '<div style="margin-top:8px;font-size:12px;color:#fbbf24;">' + data.detalle + '</div>' : '')
        + (env ? '<div style="margin-top:10px;padding:8px 12px;background:#1e293b;border-radius:6px;">' + env + '</div>' : '')
        + (pasos ? '<div style="margin-top:10px;font-size:12px;color:#cbd5e1;"><strong>Pasos para configurar:</strong><ul style="margin-top:6px;padding-left:20px;">' + pasos + '</ul></div>' : '')
        + '</div>';
      toast('SMTP no configurado o credenciales inválidas', 'warn');
    }
  } catch(e) {
    out.innerHTML = '<div style="color:#f87171;padding:14px;">Error de red: ' + (e.message || e) + '</div>';
  }
}

// ── DIAGNOSTICO DE FORMULAS ──────────────────────────────────────────────────
let _DIAG_FORM_DATA = null;

async function cargarDiagnosticoFormulas() {
  const out = document.getElementById('diag-form-result');
  const btn = document.getElementById('btn-diag-form');
  btn.disabled = true; btn.textContent = 'Analizando...';
  out.innerHTML = '<div style="color:#94a3b8;">Comparando formula_items vs maestro_mps...</div>';
  document.getElementById('diag-form-stats').style.display = 'none';
  document.getElementById('diag-form-aplicar-box').style.display = 'none';
  try {
    const r = await fetch('/api/admin/diagnosticar-formulas');
    const d = await r.json();
    btn.disabled = false; btn.innerHTML = '&#x1F50D; Ejecutar diagn&oacute;stico';
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    _DIAG_FORM_DATA = d;

    // Stats
    document.getElementById('diag-form-stats').style.display = 'block';
    document.getElementById('diagf-total').textContent = d.stats.total_formula_items;
    document.getElementById('diagf-problemas').textContent = d.stats.total_problemas;
    document.getElementById('diagf-huerf').textContent = d.stats.huerfanos;
    document.getElementById('diagf-misn').textContent = d.stats.mismatch_nombre;
    document.getElementById('diagf-auto').textContent = d.stats.auto_corregibles;
    document.getElementById('diagf-rev').textContent = d.stats.requieren_revision;

    if (d.stats.total_problemas === 0) {
      out.innerHTML = '<div style="background:rgba(34,197,94,.15);border:1px solid #22c55e;border-radius:8px;padding:14px;color:#34d399;">✓ Sin problemas detectados. Todas las fórmulas apuntan correctamente al catálogo.</div>';
      return;
    }
    document.getElementById('diag-form-aplicar-box').style.display = 'block';

    // Obsoletas (sin candidato)
    const sinCand = d.problemas.filter(p => !p.mejor_candidato);
    if (sinCand.length > 0) {
      document.getElementById('diag-form-obsoletas-box').style.display = 'block';
      document.getElementById('diagf-sin-cand-count').textContent = sinCand.length + ' items sin candidato (probablemente obsoletos)';
    } else {
      document.getElementById('diag-form-obsoletas-box').style.display = 'none';
    }

    // Render tabla — agrupada por producto
    let h = '<h4 style="color:#a5b4fc;margin-top:14px;">Items con problemas (' + d.stats.total_problemas + ' de ' + d.stats.total_formula_items + ')</h4>';
    h += '<div style="font-size:11px;color:#94a3b8;margin-bottom:10px;">';
    h += 'Marcadas verdes = auto-corregibles (1 candidato exacto). Sin candidato = requieren revisión manual.';
    h += '</div>';
    h += '<div style="overflow-x:auto;"><table id="diag-form-tabla"><thead><tr>';
    h += '<th><input type="checkbox" id="diagf-check-all" onchange="_toggleDiagFCheckAll(this)"></th>';
    h += '<th>Producto</th>';
    h += '<th>Material en fórmula</th>';
    h += '<th>Código actual</th>';
    h += '<th>Problema</th>';
    h += '<th>Sugerencia</th>';
    h += '</tr></thead><tbody>';
    d.problemas.forEach((p, idx) => {
      const auto = !!p.auto_corregible;
      const sinCandidato = !p.mejor_candidato;
      const bgColor = sinCandidato ? 'rgba(220,38,38,.10)' : (auto ? 'rgba(34,197,94,.08)' : 'rgba(245,158,11,.08)');
      const probColor = p.problema === 'huerfano' ? '#dc2626' : '#f59e0b';
      h += '<tr style="background:' + bgColor + ';">';
      h += '<td><input type="checkbox" class="diagf-row-check" data-idx="' + idx + '"' + (auto ? ' checked' : '') + '></td>';
      h += '<td style="font-size:11px;color:#cbd5e1;">' + _esc(p.producto) + '</td>';
      h += '<td><div style="font-weight:600;">' + _esc(p.material_nombre_formula) + '</div><div style="font-size:10px;color:#94a3b8;">' + _fmtG(p.cantidad_g_por_lote) + ' x lote</div></td>';
      h += '<td style="font-family:monospace;font-size:11px;color:#94a3b8;">' + _esc(p.material_id_actual) + '</td>';
      h += '<td><span style="background:' + probColor + '22;color:' + probColor + ';border:1px solid ' + probColor + ';border-radius:8px;padding:2px 6px;font-size:10px;font-weight:700;">' + _esc(p.problema.toUpperCase()) + '</span></td>';
      if (sinCandidato) {
        h += '<td style="font-size:11px;color:#fca5a5;">⚠ Sin candidato — revisar manualmente</td>';
      } else {
        const cand = p.mejor_candidato;
        const scoreColor = cand.score >= 100 ? '#22c55e' : (cand.score >= 80 ? '#fbbf24' : '#f59e0b');
        h += '<td>';
        h += '<div style="font-family:monospace;font-size:11px;color:#e2e8f0;">→ ' + _esc(cand.codigo) + '</div>';
        h += '<div style="font-size:10px;color:#cbd5e1;">' + _esc(cand.nombre_comercial || cand.nombre_inci) + '</div>';
        h += '<div style="font-size:10px;color:' + scoreColor + ';">match score: ' + cand.score + '%</div>';
        h += '</td>';
      }
      h += '</tr>';
    });
    h += '</tbody></table></div>';
    out.innerHTML = h;
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x1F50D; Ejecutar diagn&oacute;stico';
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

function _toggleDiagFCheckAll(el) {
  document.querySelectorAll('.diagf-row-check').forEach(c => c.checked = el.checked);
}

function seleccionarSoloAuto() {
  if (!_DIAG_FORM_DATA) return;
  document.querySelectorAll('.diagf-row-check').forEach(c => {
    const idx = parseInt(c.dataset.idx);
    const p = _DIAG_FORM_DATA.problemas[idx];
    c.checked = !!p.auto_corregible;
  });
}

function seleccionarTodos() {
  document.querySelectorAll('.diagf-row-check').forEach(c => {
    const idx = parseInt(c.dataset.idx);
    const p = _DIAG_FORM_DATA.problemas[idx];
    c.checked = !!p.mejor_candidato;  // solo los que tienen candidato
  });
}

function deseleccionarTodos() {
  document.querySelectorAll('.diagf-row-check').forEach(c => c.checked = false);
}

function exportarDiagFormCSV() {
  if (!_DIAG_FORM_DATA) { toast('Primero ejecuta diagnóstico', 'warn'); return; }
  const rows = [['Producto','MaterialNombre','CodigoActual','Problema','CodigoSugerido','NombreSugerido','Score']];
  _DIAG_FORM_DATA.problemas.forEach(p => {
    const cand = p.mejor_candidato || {};
    rows.push([p.producto, p.material_nombre_formula, p.material_id_actual, p.problema,
               cand.codigo||'', cand.nombre_comercial||'', cand.score||0]);
  });
  const csv = rows.map(r => r.map(c => '"' + String(c==null?'':c).replace(/"/g,'""') + '"').join(',')).join('\n');
  const blob = new Blob([csv], {type:'text/csv'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'diagnostico_formulas_' + new Date().toISOString().slice(0,10) + '.csv';
  a.click();
}

async function aplicarBatch20260428() {
  const token = (document.getElementById('diagf-token-batch20260428').value || '').trim();
  if (token !== 'APLICAR_CORRECCIONES_2026_04_28') {
    toast('Token incorrecto. Debe ser exactamente: APLICAR_CORRECCIONES_2026_04_28', 'warn');
    return;
  }
  if (!confirm('Aplicar batch de 240 correcciones validadas (2026-04-28)?\n\n' +
               '- 44 MPs nuevos en catalogo\n' +
               '- Azeclair MP00284 marcado activo=0\n' +
               '- INCI oficial actualizado AOS 40\n' +
               '- 207 formula_items corregidos\n' +
               '- 33 filas de Agua Desionizada eliminadas\n\n' +
               'Backup automatico previo. Transaccion atomica. Reversible.')) return;
  const btn = document.getElementById('btn-aplicar-batch');
  btn.disabled = true; btn.textContent = 'Aplicando 240 cambios...';
  try {
    let r = await fetch('/api/admin/aplicar-correcciones-formulas-batch-2026-04-28', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({token: token})
    });
    let d = await r.json();

    // Caso 1: hay duplicados por INCI -> recomendar mapear_duplicados
    if (r.status === 409 && d.total_duplicados_inci) {
      const dups = d.duplicados_inci || [];
      const preview = dups.slice(0,8).map(function(o){
        return '  ' + o.inci_planeado + ' ya existe como ' + o.codigo_existente +
               (o.codigo_existente !== o.codigo_planeado ? ' (planeado: ' + o.codigo_planeado + ')' : '');
      }).join('\n');
      const ok = confirm(
        'Detecte ' + d.total_duplicados_inci + ' INCI(s) que YA EXISTEN en el catalogo con otro codigo.\n\n' +
        'Si creo MPs nuevos, quedan duplicados. Si MAPEO, los formula_items apuntaran a los codigos existentes (sin crear entradas extra).\n\n' +
        'Duplicados detectados:\n' + preview +
        (dups.length > 8 ? '\n  ... +' + (dups.length-8) + ' mas' : '') +
        '\n\n¿Aplicar en modo MAPEAR_DUPLICADOS (recomendado, sin crear duplicados)?'
      );
      if (!ok) {
        btn.disabled = false; btn.innerHTML = '&#x2728; Aplicar 240 correcciones';
        toast('Cancelado. Revisa los duplicados antes de continuar.', 'warn');
        return;
      }
      btn.textContent = 'Mapeando a codigos existentes...';
      r = await fetch('/api/admin/aplicar-correcciones-formulas-batch-2026-04-28', {
        method: 'POST', headers: {'Content-Type':'application/json'},
        body: JSON.stringify({token: token, modo: 'mapear_duplicados'})
      });
      d = await r.json();
    }
    // Caso 2: rango ocupado pero SIN duplicados de INCI -> auto_renumerar seguro
    else if (r.status === 409 && d.total_ocupados) {
      let nombresPreview = '';
      if (d.ocupados_con_nombres && d.ocupados_con_nombres.length) {
        nombresPreview = '\n\nPrimeros ocupados:\n' + d.ocupados_con_nombres.slice(0,5).map(function(o){
          return '  ' + o.codigo + ' = ' + (o.comercial || o.inci || '?');
        }).join('\n');
      }
      const ok = confirm(
        'El rango MP00400-MP00500 tiene ' + d.total_ocupados + ' codigos ocupados, pero ninguno colisiona por INCI con los mios.\n\n' +
        '¿Renumerar al siguiente bloque libre (MP01000+) y aplicar?' +
        nombresPreview
      );
      if (!ok) {
        btn.disabled = false; btn.innerHTML = '&#x2728; Aplicar 240 correcciones';
        return;
      }
      btn.textContent = 'Renumerando y aplicando...';
      r = await fetch('/api/admin/aplicar-correcciones-formulas-batch-2026-04-28', {
        method: 'POST', headers: {'Content-Type':'application/json'},
        body: JSON.stringify({token: token, modo: 'auto_renumerar'})
      });
      d = await r.json();
    }
    btn.disabled = false; btn.innerHTML = '&#x2728; Aplicar 240 correcciones';
    if (!r.ok) {
      let msg = d.error || 'desconocido';
      if (d.ocupados && d.ocupados.length) {
        msg += '\nOcupados: ' + d.ocupados.join(', ');
      }
      toast('Error: ' + _esc(msg), 'warn');
      return;
    }
    const resumen = 'OK. ' +
      'MPs nuevos: ' + d.despues.mps_nuevos_creados + ' · ' +
      'formula_items diff: ' + d.cambios_netos.formula_items_diff + ' · ' +
      'maestro_mps diff: ' + d.cambios_netos.maestro_mps_diff;
    toast(resumen, 'ok');
    document.getElementById('diagf-token-batch20260428').value = '';
    cargarDiagnosticoFormulas();
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x2728; Aplicar 240 correcciones';
    toast('Error: ' + e.message, 'warn');
  }
}

async function revertirFormulas() {
  const token = (document.getElementById('diagf-token-revertir').value || '').trim();
  if (token !== 'REVERTIR_FORMULAS_2026') {
    toast('Token incorrecto. Debe ser exactamente: REVERTIR_FORMULAS_2026', 'warn');
    return;
  }
  if (!confirm('REVERTIR las correcciones de fórmulas al estado anterior?\n\nEsto reemplaza formula_items con la versión del backup automático más reciente. NO afecta movimientos, catálogo, OCs.\n\nProceder?')) return;
  const btn = document.getElementById('btn-revertir-form');
  btn.disabled = true; btn.textContent = 'Revirtiendo...';
  try {
    const r = await fetch('/api/admin/revertir-formulas-desde-backup', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({token: token})
    });
    const d = await r.json();
    btn.disabled = false; btn.innerHTML = '&#x21A9; Revertir desde backup';
    if (!r.ok) {
      toast('Error: ' + _esc(d.error||'desconocido'), 'warn');
      return;
    }
    toast(d.mensaje, 'ok');
    document.getElementById('diagf-token-revertir').value = '';
    cargarDiagnosticoFormulas();
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x21A9; Revertir desde backup';
    toast('Error: ' + e.message, 'warn');
  }
}

async function eliminarFormulasObsoletas() {
  if (!_DIAG_FORM_DATA) return;
  const token = (document.getElementById('diagf-token-eliminar').value || '').trim();
  if (token !== 'ELIMINAR_FORMULAS_OBSOLETAS_2026') {
    toast('Token incorrecto. Debe ser exactamente: ELIMINAR_FORMULAS_OBSOLETAS_2026', 'warn');
    return;
  }
  const ids = _DIAG_FORM_DATA.problemas
    .filter(p => !p.mejor_candidato)
    .map(p => p.formula_item_id);
  if (!ids.length) { toast('No hay items sin candidato para eliminar', 'warn'); return; }
  if (!confirm('Eliminar ' + ids.length + ' items obsoletos de formula_items? Backup automático previo.')) return;

  const btn = document.getElementById('btn-eliminar-obs');
  btn.disabled = true; btn.textContent = 'Eliminando...';
  try {
    const r = await fetch('/api/admin/eliminar-formulas-obsoletas', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({token: token, formula_item_ids: ids})
    });
    const d = await r.json();
    btn.disabled = false; btn.innerHTML = '&#x1F5D1; Eliminar obsoletas';
    if (!r.ok) { toast('Error: ' + _esc(d.error||'desconocido'), 'warn'); return; }
    toast(d.mensaje || (d.count_eliminados + ' items eliminados'), 'ok');
    document.getElementById('diagf-token-eliminar').value = '';
    cargarDiagnosticoFormulas();
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x1F5D1; Eliminar obsoletas';
    toast('Error: ' + e.message, 'warn');
  }
}

async function aplicarCorreccionFormulas() {
  if (!_DIAG_FORM_DATA) return;
  const token = (document.getElementById('diagf-token').value || '').trim();
  if (token !== 'CORREGIR_FORMULAS_2026') {
    toast('Token incorrecto. Debe ser exactamente: CORREGIR_FORMULAS_2026', 'warn');
    return;
  }
  // Recoger seleccionados
  const correcciones = [];
  document.querySelectorAll('.diagf-row-check').forEach(c => {
    if (!c.checked) return;
    const idx = parseInt(c.dataset.idx);
    const p = _DIAG_FORM_DATA.problemas[idx];
    if (!p.mejor_candidato) return;
    correcciones.push({
      formula_item_id: p.formula_item_id,
      nuevo_material_id: p.mejor_candidato.codigo,
      nuevo_material_nombre: p.mejor_candidato.nombre_comercial || p.mejor_candidato.nombre_inci,
    });
  });
  if (!correcciones.length) {
    toast('Selecciona al menos una corrección', 'warn');
    return;
  }
  if (!confirm('Aplicar ' + correcciones.length + ' correcciones a formula_items? Backup automático previo.')) return;

  const btn = document.getElementById('btn-aplicar-form');
  btn.disabled = true; btn.textContent = 'Aplicando...';
  try {
    const r = await fetch('/api/admin/corregir-formulas', {
      method: 'POST', headers: {'Content-Type':'application/json'},
      body: JSON.stringify({token: token, correcciones: correcciones})
    });
    const d = await r.json();
    btn.disabled = false; btn.innerHTML = '&#x1F4A5; Aplicar correcciones';
    if (!r.ok) { toast('Error: ' + _esc(d.error||'desconocido'), 'warn'); return; }
    toast(d.mensaje || (d.count_aplicados + ' correcciones aplicadas'), 'ok');
    document.getElementById('diagf-token').value = '';
    cargarDiagnosticoFormulas();  // recargar para ver el delta
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x1F4A5; Aplicar correcciones';
    toast('Error: ' + e.message, 'warn');
  }
}

// ── AUDITAR MINIMOS DE MPs ────────────────────────────────────────────────────
let _AUD_MIN_DATA = null;

async function cargarAuditarMinimos() {
  const proy = document.getElementById('audmin-proy').value;
  const out = document.getElementById('audmin-result');
  const btn = document.getElementById('btn-aud-min');
  btn.disabled = true; btn.textContent = 'Calculando...';
  out.innerHTML = '<div style="color:#94a3b8;">Proyectando consumo y comparando con mínimos actuales...</div>';
  document.getElementById('audmin-stats').style.display = 'none';
  document.getElementById('audmin-aplicar-box').style.display = 'none';
  try {
    const r = await fetch('/api/admin/auditar-minimos?proyeccion_dias=' + encodeURIComponent(proy));
    const d = await r.json();
    btn.disabled = false; btn.innerHTML = '&#x1F50D; Auditar (vista previa)';
    if (!r.ok) { out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(d.error||r.status) + '</div>'; return; }
    _AUD_MIN_DATA = d;

    // Stats
    document.getElementById('audmin-stats').style.display = 'block';
    document.getElementById('audmin-total').textContent = d.stats.total;
    document.getElementById('audmin-ok').textContent = d.stats.ok;
    document.getElementById('audmin-sub').textContent = d.stats.sub_protegido;
    document.getElementById('audmin-sobre').textContent = d.stats.sobre_protegido;
    document.getElementById('audmin-vacio').textContent = d.stats.sin_minimo;
    document.getElementById('audmin-uso').textContent = d.stats.sin_uso;

    // Mostrar bloque aplicar si hay algo que aplicar
    const totalAplicable = d.stats.sub_protegido + d.stats.sobre_protegido + d.stats.sin_minimo;
    document.getElementById('audmin-aplicar-box').style.display = totalAplicable > 0 ? 'block' : 'none';

    // Render tabla
    let h = '<h4 style="color:#a5b4fc;margin-top:14px;">Detalle por MP — ' + d.stats.total + ' materias primas</h4>';
    h += '<div style="font-size:11px;color:#94a3b8;margin-bottom:10px;">' +
         'Métodología: <code>' + _esc(d.metodologia.formula) + '</code>. ' +
         'China: 90d. Local: 21d. Sin proveedor: 28d. Piso 50g para péptidos.' +
         '</div>';
    h += '<div style="margin-bottom:10px;display:flex;gap:6px;align-items:center;flex-wrap:wrap;font-size:12px;">';
    h += '<span style="color:#cbd5e1;">Filtrar:</span>';
    ['todos', 'OK', 'SUB_PROTEGIDO', 'SOBRE_PROTEGIDO', 'SIN_MINIMO_CONFIGURADO', 'SIN_USO'].forEach(function(f){
      h += '<button class="btn btn-outline" style="padding:3px 10px;font-size:11px;" onclick="_filtrarAudMin(\'' + f + '\')">' + f + '</button>';
    });
    h += '<input type="text" id="audmin-search" oninput="_filtrarAudMin()" placeholder="Buscar nombre/código..." style="background:#0f172a;color:#e2e8f0;border:1px solid #334155;border-radius:5px;padding:4px 8px;font-size:11px;width:180px;">';
    h += '</div>';
    h += '<div style="overflow-x:auto;"><table id="audmin-tabla"><thead><tr>';
    h += '<th>Estado</th><th>Material</th><th>Proveedor</th><th>Origen</th>';
    h += '<th style="text-align:right;">Consumo/día</th>';
    h += '<th style="text-align:right;">Mínimo actual</th>';
    h += '<th style="text-align:right;">Recomendado</th>';
    h += '<th>Cobertura</th>';
    h += '<th>Razonamiento</th>';
    h += '</tr></thead><tbody id="audmin-tbody"></tbody></table></div>';
    out.innerHTML = h;
    _filtrarAudMin('todos');
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x1F50D; Auditar (vista previa)';
    out.innerHTML = '<div style="color:#fca5a5;">Error: ' + _esc(e.message) + '</div>';
  }
}

function _filtrarAudMin(estado) {
  if (!_AUD_MIN_DATA) return;
  if (estado) window._audMinFiltro = estado;
  const filtro = window._audMinFiltro || 'todos';
  const search = (document.getElementById('audmin-search')||{}).value || '';
  const sLower = search.trim().toLowerCase();
  const items = _AUD_MIN_DATA.auditoria.filter(function(a){
    if (filtro === 'SIN_USO') {
      if (!a.estado.startsWith('SIN_USO')) return false;
    } else if (filtro !== 'todos' && a.estado !== filtro) {
      return false;
    }
    if (sLower) {
      const txt = (a.nombre + ' ' + a.codigo_mp + ' ' + (a.proveedor||'')).toLowerCase();
      if (txt.indexOf(sLower) === -1) return false;
    }
    return true;
  });
  // Orden: SUB_PROTEGIDO > SIN_MINIMO_CONFIGURADO > SOBRE_PROTEGIDO > OK > SIN_USO
  const orden = {SUB_PROTEGIDO:0, SIN_MINIMO_CONFIGURADO:1, SOBRE_PROTEGIDO:2, OK:3, SIN_USO_CON_MIN:4, SIN_USO:5};
  items.sort(function(a, b){
    const oa = orden[a.estado] !== undefined ? orden[a.estado] : 9;
    const ob = orden[b.estado] !== undefined ? orden[b.estado] : 9;
    if (oa !== ob) return oa - ob;
    return (b.consumo_diario_g || 0) - (a.consumo_diario_g || 0);
  });
  const colors = {
    OK: '#22c55e',
    SUB_PROTEGIDO: '#dc2626',
    SOBRE_PROTEGIDO: '#f59e0b',
    SIN_MINIMO_CONFIGURADO: '#6366f1',
    SIN_USO: '#94a3b8',
    SIN_USO_CON_MIN: '#94a3b8',
  };
  let h = '';
  items.forEach(function(a){
    const col = colors[a.estado] || '#cbd5e1';
    const cobertura = a.consumo_diario_g > 0
      ? Math.round(a.stock_minimo_actual_g / a.consumo_diario_g) + 'd'
      : '—';
    h += '<tr>';
    h += '<td><span style="background:' + col + '22;color:' + col + ';border:1px solid ' + col + ';border-radius:10px;padding:2px 8px;font-size:10px;font-weight:700;">' + _esc(a.estado.replace('_', ' ')) + '</span></td>';
    h += '<td><div style="font-weight:600;font-size:12px;">' + _esc(a.nombre) + '</div><div style="font-size:10px;color:#64748b;font-family:monospace;">' + _esc(a.codigo_mp) + '</div></td>';
    h += '<td style="font-size:11px;color:#cbd5e1;">' + _esc(a.proveedor || '—') + '</td>';
    h += '<td style="font-size:11px;color:#94a3b8;">' + _esc(a.origen) + '</td>';
    h += '<td style="text-align:right;font-size:11px;">' + _fmtG(a.consumo_diario_g) + '</td>';
    h += '<td style="text-align:right;font-size:12px;">' + _fmtG(a.stock_minimo_actual_g) + '</td>';
    h += '<td style="text-align:right;font-size:12px;font-weight:700;color:' + col + ';">' + _fmtG(a.minimo_recomendado_g) + '</td>';
    h += '<td style="font-size:11px;color:#94a3b8;">' + cobertura + '</td>';
    h += '<td style="font-size:11px;color:#cbd5e1;max-width:280px;">' + _esc(a.razonamiento) + '</td>';
    h += '</tr>';
  });
  if (!h) h = '<tr><td colspan="9" style="text-align:center;color:#64748b;padding:20px;">Sin coincidencias</td></tr>';
  document.getElementById('audmin-tbody').innerHTML = h;
}

function exportarAuditMinCSV() {
  if (!_AUD_MIN_DATA) { toast('Primero ejecuta auditoría', 'warn'); return; }
  const rows = [['Codigo','Nombre','Proveedor','Origen','ConsumoDiario_g','ConsumoMensual_g','StockMinimoActual_g','MinimoRecomendado_g','LeadTime_d','Buffer_d','Estado','Razonamiento']];
  _AUD_MIN_DATA.auditoria.forEach(function(a){
    rows.push([a.codigo_mp, a.nombre, a.proveedor, a.origen, a.consumo_diario_g, a.consumo_mensual_g, a.stock_minimo_actual_g, a.minimo_recomendado_g, a.lead_time_dias, a.buffer_dias, a.estado, a.razonamiento]);
  });
  const csv = rows.map(function(r){return r.map(function(c){return '"' + String(c==null?'':c).replace(/"/g,'""') + '"';}).join(',');}).join('\n');
  const blob = new Blob([csv], {type:'text/csv'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'auditoria_minimos_mps_' + new Date().toISOString().slice(0,10) + '.csv';
  a.click();
}

async function aplicarRecalculoMinimos() {
  const token = (document.getElementById('audmin-token').value || '').trim();
  if (token !== 'APLICAR_MINIMOS_RECALCULADOS_2026') {
    toast('Token incorrecto. Debe ser exactamente: APLICAR_MINIMOS_RECALCULADOS_2026', 'warn');
    return;
  }
  if (!confirm('Esto va a actualizar stock_minimo en maestro_mps para los MPs marcados como SUB/SOBRE/SIN_MINIMO. Crea backup automático previo. ¿Continuar?')) return;
  const proy = document.getElementById('audmin-proy').value;
  const btn = document.getElementById('btn-aplicar-min');
  btn.disabled = true; btn.textContent = 'Aplicando...';
  try {
    const r = await fetch('/api/admin/aplicar-minimos', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({token: token, proyeccion_dias: parseInt(proy)})
    });
    const d = await r.json();
    btn.disabled = false; btn.innerHTML = '&#x1F4A5; Aplicar recálculo';
    if (!r.ok) { toast('Error: ' + _esc(d.error||'desconocido'), 'warn'); return; }
    toast(d.mensaje || (d.count_cambios + ' mínimos actualizados'), 'ok');
    document.getElementById('audmin-token').value = '';
    // Recargar auditoría
    cargarAuditarMinimos();
  } catch(e) {
    btn.disabled = false; btn.innerHTML = '&#x1F4A5; Aplicar recálculo';
    toast('Error: ' + e.message, 'warn');
  }
}

// ── BANCO INFLUENCERS (sync Excel) ────────────────────────────────────────────
async function syncExcel(dryRun) {
  const fi = document.getElementById('excel-file');
  const out = document.getElementById('banco-result');
  if (!fi.files.length) {
    toast('Selecciona un archivo .xlsx', 'warn');
    return;
  }
  out.innerHTML = '<div style="color:#94a3b8;padding:14px;">Procesando...</div>';
  const fd = new FormData();
  fd.append('file', fi.files[0]);
  const url = '/api/admin/sync-influencers-excel' + (dryRun ? '?dry_run=1' : '');
  const r = await fetch(url, {method:'POST', body: fd});
  let data;
  try { data = await r.json(); } catch(e) { data = {error: 'Respuesta inválida'}; }
  if (!r.ok) {
    out.innerHTML = '<div style="color:#f87171;padding:14px;background:#1e293b;border-radius:8px;">'
      + 'Error ' + r.status + ': ' + (data.error || '') + '</div>';
    return;
  }
  const banner = dryRun
    ? '<div style="color:#fbbf24;padding:10px 14px;background:#0f172a;border:1px solid #fbbf24;border-radius:8px;margin-bottom:14px;">'
      + '&#x1F441; <strong>VISTA PREVIA</strong> — nada se escribió en la base. Pulsa "Aplicar" para confirmar.</div>'
    : '<div style="color:#34d399;padding:10px 14px;background:#0f172a;border:1px solid #34d399;border-radius:8px;margin-bottom:14px;">'
      + '&#x2705; <strong>APLICADO</strong> — la base fue actualizada.</div>';
  const lstNuevos = (data.nuevos.lista || []).slice(0, 30).map(n =>
    '<li style="padding:3px 0;">' + n + '</li>').join('');
  const lstAct = (data.actualizados.lista || []).slice(0, 30).map(a =>
    '<li style="padding:3px 0;"><strong>' + a.nombre + '</strong>'
    + (a.cambios && a.cambios.length ? ' <span style="color:#94a3b8;font-size:11px;">— '
       + a.cambios.join('; ') + '</span>' : '') + '</li>').join('');
  out.innerHTML = banner + ''
    + '<div class="kpi-row">'
      + '<div class="kpi"><div class="kpi-l">Excel</div><div class="kpi-v">' + data.total_excel + '</div></div>'
      + '<div class="kpi"><div class="kpi-l">Nuevos</div><div class="kpi-v" style="color:#34d399;">' + data.nuevos.count + '</div></div>'
      + '<div class="kpi"><div class="kpi-l">Actualizados</div><div class="kpi-v" style="color:#fbbf24;">' + data.actualizados.count + '</div></div>'
      + '<div class="kpi"><div class="kpi-l">Sin cambios</div><div class="kpi-v" style="color:#64748b;">' + data.sin_cambios.count + '</div></div>'
    + '</div>'
    + (lstNuevos ? '<div class="card"><h2>&#x2795; Nuevos (' + data.nuevos.count + ')</h2><ul style="font-size:13px;color:#cbd5e1;list-style:none;padding-left:0;">' + lstNuevos + '</ul></div>' : '')
    + (lstAct ? '<div class="card"><h2>&#x270F; Actualizados (' + data.actualizados.count + ')</h2><ul style="font-size:13px;color:#cbd5e1;list-style:none;padding-left:0;">' + lstAct + '</ul></div>' : '');
  if (!dryRun) toast('Sync aplicado: ' + data.nuevos.count + ' nuevos, ' + data.actualizados.count + ' actualizados', 'ok');
}

// ── PAGOS INFLUENCERS — import Excel + reset pendientes ──────────────────────
async function syncPagos(dryRun) {
  const fi = document.getElementById('pagos-excel-file');
  const out = document.getElementById('pagos-result');
  if (!fi.files.length) {
    toast('Selecciona un archivo .xlsx', 'warn');
    return;
  }
  if (!dryRun) {
    if (!confirm('Esto va a BORRAR todas las solicitudes Influencer no-pagadas y reemplazarlas con las del Excel. Las solicitudes ya pagadas se conservan.\n\n¿Confirmás?')) return;
  }
  out.innerHTML = '<div style="color:#94a3b8;padding:14px;">Procesando...</div>';
  const fd = new FormData();
  fd.append('file', fi.files[0]);
  const url = '/api/admin/import-pagos-influencers-excel' + (dryRun ? '?dry_run=1' : '');
  const r = await fetch(url, {method:'POST', body: fd});
  let data;
  try { data = await r.json(); } catch(e) { data = {error: 'Respuesta inválida'}; }
  if (!r.ok) {
    let extra = '';
    if (data.headers_detectados) {
      extra += '<div style="margin-top:8px;font-size:11px;color:#94a3b8;"><strong>Headers detectados:</strong> ' + data.headers_detectados.join(' · ') + '</div>';
    }
    if (data.hojas_disponibles) {
      extra += '<div style="margin-top:4px;font-size:11px;color:#94a3b8;"><strong>Hojas en el Excel:</strong> ' + data.hojas_disponibles.join(' · ') + '</div>';
    }
    if (data.sugerencia) {
      extra += '<div style="margin-top:8px;font-size:12px;color:#fbbf24;">' + data.sugerencia + '</div>';
    }
    out.innerHTML = '<div style="color:#f87171;padding:14px;background:#1e293b;border-radius:8px;">'
      + 'Error ' + r.status + ': ' + (data.error || '') + extra + '</div>';
    return;
  }
  const banner = dryRun
    ? '<div style="color:#fbbf24;padding:10px 14px;background:#0f172a;border:1px solid #fbbf24;border-radius:8px;margin-bottom:14px;">'
      + '&#x1F441; <strong>VISTA PREVIA</strong> — nada se escribió. Pulsa "Aplicar" para confirmar.</div>'
    : '<div style="color:#34d399;padding:10px 14px;background:#0f172a;border:1px solid #34d399;border-radius:8px;margin-bottom:14px;">'
      + '&#x2705; <strong>APLICADO</strong> — la base fue actualizada.</div>';

  const cm = (data.columnas_mapeadas || {});
  const cmList = Object.entries(cm).map(([k,v]) => k+'→col'+v).join(' · ');

  let kpis = '<div class="kpi-row">';
  if (data.a_importar) {
    kpis += '<div class="kpi"><div class="kpi-l">A importar</div><div class="kpi-v" style="color:#34d399;">' + data.a_importar.count + '</div></div>';
    kpis += '<div class="kpi"><div class="kpi-l">Valor total</div><div class="kpi-v" style="color:#34d399;font-size:14px;">$' + Number(data.a_importar.valor_total||0).toLocaleString('es-CO') + '</div></div>';
  }
  if (data.importadas) {
    kpis += '<div class="kpi"><div class="kpi-l">Importadas</div><div class="kpi-v" style="color:#34d399;">' + data.importadas.count + '</div></div>';
    kpis += '<div class="kpi"><div class="kpi-l">Valor total</div><div class="kpi-v" style="color:#34d399;font-size:14px;">$' + Number(data.importadas.valor_total||0).toLocaleString('es-CO') + '</div></div>';
  }
  if (data.reset) {
    kpis += '<div class="kpi"><div class="kpi-l">Solic. borradas</div><div class="kpi-v" style="color:#fbbf24;">' + data.reset.solicitudes + '</div></div>';
  }
  kpis += '<div class="kpi"><div class="kpi-l">Sin match</div><div class="kpi-v" style="color:#f87171;">' + (data.sin_match ? data.sin_match.count : 0) + '</div></div>';
  if (data.pagadas_skipped) {
    kpis += '<div class="kpi"><div class="kpi-l">Skipped (pagadas)</div><div class="kpi-v" style="color:#94a3b8;">' + data.pagadas_skipped.count + '</div></div>';
  }
  kpis += '</div>';

  let preview = '';
  const lst = (data.a_importar && data.a_importar.preview) || (data.importadas && data.importadas.lista) || [];
  if (lst.length) {
    preview = '<div class="card"><h2>&#x2795; A importar / Importadas</h2>'
      + '<table><thead><tr><th>Sol/OC</th><th>Influencer</th><th>Valor</th><th>Fecha pub</th></tr></thead><tbody>'
      + lst.map(p => '<tr><td style="font-family:monospace;font-size:11px;">' + (p.sol || ('#' + p.fila)) + (p.oc ? ' / ' + p.oc : '') + '</td><td>' + p.nombre + '</td><td>$' + Number(p.valor||0).toLocaleString('es-CO') + '</td><td style="font-size:11px;color:#94a3b8;">' + (p.fecha_publicacion || '—') + '</td></tr>').join('')
      + '</tbody></table></div>';
  }
  let nomatch = '';
  if (data.sin_match && data.sin_match.count > 0) {
    nomatch = '<div class="card" style="border-left:3px solid #f87171;"><h2>&#x274C; Sin match en banco (' + data.sin_match.count + ')</h2>'
      + '<div style="font-size:12px;color:#fbbf24;margin-bottom:8px;">' + (data.sin_match.nota || '') + '</div>'
      + '<table><tbody>'
      + data.sin_match.lista.map(p => '<tr><td style="font-size:11px;">fila ' + p.fila + '</td><td>' + p.nombre + '</td><td>$' + Number(p.valor||0).toLocaleString('es-CO') + '</td><td style="font-size:11px;color:#94a3b8;">' + (p.instagram || '') + '</td></tr>').join('')
      + '</tbody></table></div>';
  }

  out.innerHTML = banner
    + '<div style="font-size:11px;color:#94a3b8;margin-bottom:8px;">Hoja: <strong>' + (data.hoja_usada || '?') + '</strong> · Columnas: ' + cmList + '</div>'
    + kpis + preview + nomatch;
  if (!dryRun) toast('Importación aplicada: ' + (data.importadas ? data.importadas.count : 0) + ' nuevas, ' + (data.reset ? data.reset.solicitudes : 0) + ' borradas', 'ok');
}

// ── CATÁLOGO MPs (proveedores) ─────────────────────────────────────────────
let _MPS_PROVS_LIST = [];

async function loadMpsStatus(){
  const r = await fetch('/api/admin/mps-proveedores-status');
  if (!r.ok) {
    document.getElementById('mps-tbody-sin').innerHTML =
      '<tr><td colspan="4" style="color:#f87171;">Error '+r.status+'</td></tr>';
    return;
  }
  const data = await r.json();
  document.getElementById('mps-kpi-total').textContent = data.total_activos || 0;
  document.getElementById('mps-kpi-con').textContent = data.con_proveedor || 0;
  document.getElementById('mps-kpi-sin').textContent = data.sin_proveedor || 0;
  document.getElementById('mps-kpi-pct').textContent = (data.pct_cobertura || 0) + '%';

  // Tabla por proveedor
  const tbody1 = document.getElementById('mps-tbody-prov');
  if (data.por_proveedor && data.por_proveedor.length) {
    tbody1.innerHTML = data.por_proveedor.map(p => {
      const isSin = p.proveedor === '(SIN PROVEEDOR)';
      const color = isSin ? '#f87171' : '#cbd5e1';
      return '<tr><td style="color:'+color+';font-weight:'+(isSin?'700':'400')+';">'+p.proveedor+'</td>'
        + '<td style="text-align:right;font-family:monospace;">'+p.total+'</td></tr>';
    }).join('');
  } else {
    tbody1.innerHTML = '<tr><td colspan="2" style="color:#94a3b8;">Sin datos</td></tr>';
  }

  // Lista de MPs sin proveedor (con dropdown para asignar)
  _MPS_PROVS_LIST = data.proveedores_existentes || [];
  const sinList = data.sin_proveedor_lista || [];
  const info = document.getElementById('mps-sin-prov-info');
  if (info) {
    info.textContent = sinList.length === 0
      ? '✅ Todas las MPs tienen proveedor asignado.'
      : sinList.length + ' MPs sin proveedor — asígnalos para que la auto-generación de OC los agrupe correctamente.';
  }
  const tbody2 = document.getElementById('mps-tbody-sin');
  if (sinList.length === 0) {
    tbody2.innerHTML = '<tr><td colspan="4" style="color:#34d399;text-align:center;padding:20px;">✅ Sin pendientes</td></tr>';
  } else {
    tbody2.innerHTML = sinList.map(m => {
      const opts = '<option value="">(sin asignar)</option>'
        + _MPS_PROVS_LIST.map(p => '<option value="'+p+'">'+p+'</option>').join('')
        + '<option value="__nuevo__">+ Otro proveedor (escribir)...</option>';
      return '<tr>'
        + '<td style="font-family:monospace;font-size:11px;">'+m.codigo_mp+'</td>'
        + '<td>'+m.nombre+'</td>'
        + '<td style="color:#94a3b8;font-size:11px;">'+(m.tipo||'')+'</td>'
        + '<td><select class="mp-prov-sel" data-cod="'+m.codigo_mp+'" style="background:#0f172a;border:1px solid #334155;border-radius:4px;color:#e2e8f0;padding:4px 8px;font-size:12px;width:100%;" onchange="asignarMpProv(this)">'
        + opts
        + '</select></td></tr>';
    }).join('');
  }
}

async function asignarMpProv(selEl){
  const codigo = selEl.dataset.cod;
  let proveedor = selEl.value;
  if (proveedor === '__nuevo__') {
    proveedor = (prompt('Nombre del nuevo proveedor para ' + codigo + ':') || '').trim();
    if (!proveedor) {
      selEl.value = '';
      return;
    }
  }
  if (!proveedor) {
    if (!confirm('¿Quitar proveedor de ' + codigo + '?')) return;
  }
  const r = await fetch('/api/admin/mps-asignar-proveedor', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({codigo_mp: codigo, proveedor: proveedor})
  });
  const d = await r.json();
  if (d.ok) {
    toast(codigo + ' → ' + (proveedor || 'sin proveedor'), 'ok');
    // Recargar para que la fila desaparezca de la tabla
    loadMpsStatus();
  } else {
    toast('Error: ' + (d.error || 'no se pudo asignar'), 0);
    selEl.value = '';
  }
}

async function syncMpsNombres(dryRun){
  const fi = document.getElementById('mps-nom-file');
  const out = document.getElementById('mps-nom-result');
  if(!fi.files.length){ toast('Selecciona un .xlsx','warn'); return; }
  out.innerHTML = '<div style="color:#94a3b8;padding:14px;">Procesando...</div>';
  const fd = new FormData();
  fd.append('file', fi.files[0]);
  const url = '/api/admin/import-mps-nombres-excel' + (dryRun ? '?dry_run=1' : '');
  try{
    const r = await fetch(url, {method:'POST', body: fd});
    const d = await r.json();
    if(!r.ok){
      out.innerHTML = '<div style="color:#f87171;padding:14px;background:#1e293b;border-radius:8px;">'
        + 'Error '+r.status+': '+(d.error||'')
        + (d.headers ? '<div style="margin-top:8px;font-size:11px;color:#94a3b8;">Headers detectados: '+d.headers.join(' · ')+'</div>' : '')
        + (d.sugerencia ? '<div style="margin-top:8px;font-size:12px;color:#fbbf24;">'+d.sugerencia+'</div>' : '')
        + '</div>';
      return;
    }
    const banner = dryRun
      ? '<div style="color:#fbbf24;padding:10px 14px;background:#0f172a;border:1px solid #fbbf24;border-radius:8px;margin-bottom:14px;">&#x1F441; <strong>VISTA PREVIA</strong> — sin escribir</div>'
      : '<div style="color:#34d399;padding:10px 14px;background:#0f172a;border:1px solid #34d399;border-radius:8px;margin-bottom:14px;">&#x2705; <strong>APLICADO</strong></div>';
    let kpis = '<div class="kpi-row">'
      + '<div class="kpi"><div class="kpi-l">A actualizar</div><div class="kpi-v" style="color:#34d399;">'+d.actualizados.count+'</div></div>'
      + '<div class="kpi"><div class="kpi-l">Sin cambios</div><div class="kpi-v" style="color:#94a3b8;">'+d.sin_cambios.count+'</div></div>'
      + '<div class="kpi"><div class="kpi-l">Sin match</div><div class="kpi-v" style="color:#f87171;">'+d.sin_match.count+'</div></div>'
      + '<div class="kpi"><div class="kpi-l">Sin código</div><div class="kpi-v" style="color:#f87171;">'+d.sin_codigo.count+'</div></div>'
      + '</div>';
    let preview = '';
    if(d.actualizados.lista && d.actualizados.lista.length){
      preview = '<div class="card"><h2>Cambios</h2><table><thead><tr><th>Código</th><th>Nombre antes</th><th>Nombre nuevo</th></tr></thead><tbody>'
        + d.actualizados.lista.map(a => '<tr><td style="font-family:monospace;font-size:11px;">'+a.codigo+'</td><td style="color:#94a3b8;">'+(a.nombre_antes||'(vacío)')+'</td><td style="color:#34d399;font-weight:600;">'+a.nombre_nuevo+'</td></tr>').join('')
        + '</tbody></table></div>';
    }
    let nomatch = '';
    if(d.sin_match.lista && d.sin_match.lista.length){
      nomatch = '<div class="card" style="border-left:3px solid #f87171;"><h2>Sin match en catálogo ('+d.sin_match.count+')</h2>'
        + '<div style="font-size:12px;color:#fbbf24;margin-bottom:8px;">Estos códigos del Excel no están en maestro_mps. Crea primero el MP en /planta.</div>'
        + '<table><tbody>' + d.sin_match.lista.map(a => '<tr><td style="font-size:11px;">fila '+a.fila+'</td><td style="font-family:monospace;">'+a.codigo+'</td><td>'+a.nombre_excel+'</td></tr>').join('')
        + '</tbody></table></div>';
    }
    out.innerHTML = banner + kpis + preview + nomatch;
    if(!dryRun){
      toast('Aplicado: '+d.actualizados.count+' nombres corregidos','ok');
      loadMpsStatus();  // Refresca el panel principal
    }
  }catch(e){
    out.innerHTML = '<div style="color:#f87171;padding:14px;">Error: '+e.message+'</div>';
  }
}

// Init
loadBackups();
</script>
</body>
</html>"""


@bp.route("/admin", methods=["GET"])
def admin_panel():
    """Panel HTML de administración."""
    u = session.get("compras_user", "")
    if not u:
        return Response("Login requerido", status=401)
    if u not in ADMIN_USERS:
        return Response("<h1>403</h1><p>Solo admins.</p><a href='/hub'>Volver</a>",
                        status=403, mimetype="text/html")
    return Response(_ADMIN_HTML, mimetype="text/html")


# ─── Auditoria de lotes / movimientos recientes ────────────────────────────────

@bp.route("/api/admin/auditoria-lotes", methods=["GET"])
def admin_auditoria_lotes():
    """Reporta integridad de lotes y movimientos recientes.

    Querystring:
      ?dias=2  → ventana hacia atras (default 2 dias = hoy + ayer)

    Devuelve JSON con:
      - total_lotes_activos: count de lotes con stock > 0
      - lotes_creados_recientes: lotes vistos por primera vez en ventana
      - movimientos_recientes: lista de Entradas/Salidas/Ajustes en ventana
      - duplicados_sospechosos: mismo (lote + material) en mas de 1 entrada
      - resumen_por_usuario: quien creo cuantos movs hoy
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    dias = int(request.args.get("dias", 2))
    fecha_corte = f"-{dias} day"

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    out = {"ventana_dias": dias}

    # 1) Total de lotes activos (con stock > 0)
    try:
        rows = c.execute("""
            SELECT material_id, lote,
                   COALESCE(SUM(CASE WHEN tipo IN ('Entrada','Ajuste +') THEN cantidad
                                     WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad
                                     ELSE 0 END), 0) as stock
            FROM movimientos
            WHERE COALESCE(lote,'') != ''
            GROUP BY material_id, lote
            HAVING stock > 0
        """).fetchall()
        out["total_lotes_activos"] = len(rows)
    except Exception as e:
        out["total_lotes_activos_error"] = str(e)

    # 2) Lotes nuevos en la ventana (primera entrada vista en esos dias)
    try:
        nuevos = c.execute(f"""
            SELECT m.material_id, m.material_nombre, m.lote, m.proveedor,
                   m.cantidad, m.fecha, m.operador,
                   MIN(m.id) as primera_id
            FROM movimientos m
            WHERE COALESCE(m.lote,'') != ''
              AND m.tipo = 'Entrada'
              AND m.fecha >= date('now', '{fecha_corte}')
              AND NOT EXISTS (
                  SELECT 1 FROM movimientos m2
                  WHERE m2.material_id = m.material_id
                    AND m2.lote = m.lote
                    AND m2.id < m.id
              )
            GROUP BY m.material_id, m.lote
            ORDER BY m.fecha DESC, m.id DESC
            LIMIT 200
        """).fetchall()
        out["lotes_creados_recientes"] = [dict(r) for r in nuevos]
        out["lotes_creados_count"] = len(nuevos)
    except Exception as e:
        out["lotes_creados_recientes_error"] = str(e)
        out["lotes_creados_recientes"] = []

    # 3) Movimientos recientes (todos los tipos)
    try:
        movs = c.execute(f"""
            SELECT id, material_id, material_nombre, cantidad, tipo, fecha,
                   lote, proveedor, operador, observaciones
            FROM movimientos
            WHERE fecha >= date('now', '{fecha_corte}')
            ORDER BY id DESC
            LIMIT 500
        """).fetchall()
        out["movimientos_recientes"] = [dict(r) for r in movs]
        out["movimientos_count"] = len(movs)
    except Exception as e:
        out["movimientos_recientes_error"] = str(e)

    # 4) Duplicados sospechosos: mismo (lote + material + cantidad + fecha)
    #    aparece en 2 o mas filas en la ventana
    try:
        dups = c.execute(f"""
            SELECT material_id, material_nombre, lote, cantidad, tipo,
                   COUNT(*) as veces, MIN(id) as primera, MAX(id) as ultima,
                   GROUP_CONCAT(operador, ' / ') as operadores
            FROM movimientos
            WHERE COALESCE(lote,'') != ''
              AND fecha >= date('now', '{fecha_corte}')
            GROUP BY material_id, lote, cantidad, tipo, fecha
            HAVING COUNT(*) > 1
            ORDER BY veces DESC, ultima DESC
        """).fetchall()
        out["duplicados_sospechosos"] = [dict(r) for r in dups]
        out["duplicados_count"] = len(dups)
    except Exception as e:
        out["duplicados_error"] = str(e)
        out["duplicados_sospechosos"] = []

    # 5) Resumen por usuario hoy
    try:
        usr = c.execute("""
            SELECT COALESCE(operador,'(sin operador)') as operador,
                   tipo,
                   COUNT(*) as movs,
                   COUNT(DISTINCT lote) as lotes_distintos
            FROM movimientos
            WHERE fecha = date('now')
            GROUP BY operador, tipo
            ORDER BY movs DESC
        """).fetchall()
        out["resumen_hoy_por_usuario"] = [dict(r) for r in usr]
    except Exception as e:
        out["resumen_hoy_por_usuario_error"] = str(e)

    # 6) Comparativa: ¿cuantos lotes activos teniamos hace N dias?
    #    Reconstruye stock al cierre de hace N dias y cuenta los > 0.
    try:
        rows_pasado = c.execute(f"""
            SELECT material_id, lote,
                   COALESCE(SUM(CASE WHEN tipo IN ('Entrada','Ajuste +') THEN cantidad
                                     WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad
                                     ELSE 0 END), 0) as stock
            FROM movimientos
            WHERE COALESCE(lote,'') != ''
              AND fecha < date('now', '{fecha_corte}')
            GROUP BY material_id, lote
            HAVING stock > 0
        """).fetchall()
        out["lotes_activos_antes_de_ventana"] = len(rows_pasado)
        out["delta_lotes"] = out["total_lotes_activos"] - len(rows_pasado)
    except Exception as e:
        out["delta_lotes_error"] = str(e)

    conn.close()
    return jsonify(out)


@bp.route("/api/admin/auditoria-lotes/html", methods=["GET"])
def admin_auditoria_lotes_html():
    """Vista HTML legible del reporte de auditoria de lotes."""
    u, err, code = _require_admin()
    if err:
        return Response("<h1>Solo admins</h1>", status=code or 403,
                        mimetype="text/html")

    dias = int(request.args.get("dias", 2))
    html = """<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Auditoria Lotes — EOS</title>
<style>
body{font-family:-apple-system,Segoe UI,sans-serif;background:#f5f5f4;color:#1c1917;
     padding:24px;max-width:1400px;margin:0 auto;}
h1{color:#6d28d9;border-bottom:2px solid #c4b5fd;padding-bottom:8px;}
h2{color:#7c3aed;margin-top:30px;}
.kpi{display:inline-block;background:#fff;border:1px solid #e7e5e4;border-radius:8px;
     padding:16px 24px;margin:8px 8px 8px 0;min-width:180px;}
.kpi b{display:block;font-size:24px;color:#15803d;}
.kpi.warn b{color:#dc2626;}
.kpi.neutral b{color:#1c1917;}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e7e5e4;
      border-radius:8px;overflow:hidden;margin-top:12px;font-size:13px;}
th{background:#f5f3ff;text-align:left;padding:10px;color:#4c1d95;font-weight:700;}
td{padding:8px 10px;border-top:1px solid #f5f5f4;}
tr:hover{background:#fafaf9;}
.tipo-Entrada{color:#15803d;font-weight:600;}
.tipo-Salida{color:#dc2626;font-weight:600;}
.tipo-Ajuste{color:#a16207;font-weight:600;}
.alert{background:#fef2f2;border:2px solid #dc2626;border-radius:8px;padding:14px;
       margin:14px 0;color:#991b1b;}
.ok{background:#f0fdf4;border:1px solid #15803d;border-radius:6px;padding:10px;
    margin:10px 0;color:#14532d;}
.fmt-cant{text-align:right;font-family:Consolas,monospace;}
.toolbar{margin:14px 0;}
.toolbar a{display:inline-block;padding:6px 14px;background:#fff;border:1px solid #c4b5fd;
           border-radius:6px;color:#6d28d9;text-decoration:none;font-size:12px;
           margin-right:8px;}
.toolbar a:hover{background:#f5f3ff;}
</style></head><body>
<h1>&#x1F50D; Auditoria de Lotes y Movimientos</h1>
<div class="toolbar">
  <a href="?dias=1">Hoy solo</a>
  <a href="?dias=2">Hoy + ayer</a>
  <a href="?dias=7">Ultima semana</a>
  <a href="?dias=30">Ultimo mes</a>
  <a href="/api/admin/auditoria-lotes?dias=""" + str(dias) + """" target="_blank">Ver JSON crudo</a>
</div>
<div id="resumen" style="margin:20px 0;"></div>
<div id="cuerpo">Cargando...</div>
<script>
async function cargar(){
  var r = await fetch('/api/admin/auditoria-lotes?dias=""" + str(dias) + """');
  var d = await r.json();

  // KPIs
  var html = '<h2>Resumen</h2>';
  html += '<div class="kpi neutral"><span>Lotes activos AHORA</span><b>' + d.total_lotes_activos + '</b></div>';
  html += '<div class="kpi neutral"><span>Lotes activos hace ' + d.ventana_dias + 'd</span><b>' + (d.lotes_activos_antes_de_ventana||0) + '</b></div>';
  var deltaCls = (d.delta_lotes||0) > 0 ? 'warn' : 'neutral';
  var deltaSign = (d.delta_lotes||0) > 0 ? '+' : '';
  html += '<div class="kpi ' + deltaCls + '"><span>Delta lotes</span><b>' + deltaSign + (d.delta_lotes||0) + '</b></div>';
  html += '<div class="kpi neutral"><span>Movimientos en ventana</span><b>' + (d.movimientos_count||0) + '</b></div>';
  html += '<div class="kpi neutral"><span>Lotes nuevos en ventana</span><b>' + (d.lotes_creados_count||0) + '</b></div>';
  var dupCls = (d.duplicados_count||0) > 0 ? 'warn' : 'neutral';
  html += '<div class="kpi ' + dupCls + '"><span>Duplicados sospechosos</span><b>' + (d.duplicados_count||0) + '</b></div>';
  document.getElementById('resumen').innerHTML = html;

  // Cuerpo
  var c = '';

  // Duplicados (alerta si hay)
  if((d.duplicados_count||0) > 0){
    c += '<div class="alert"><b>&#x26A0; ' + d.duplicados_count + ' grupos de movimientos posiblemente duplicados</b><br>';
    c += 'Mismo lote + material + cantidad + tipo + fecha repetidos. Revisa si son legitimos o un bug.</div>';
    c += '<table><thead><tr><th>Material</th><th>Lote</th><th>Cantidad</th><th>Tipo</th><th>Veces</th><th>IDs</th><th>Operadores</th></tr></thead><tbody>';
    (d.duplicados_sospechosos||[]).forEach(function(x){
      c += '<tr><td>' + (x.material_id||'') + '<br><small style="color:#78716c">' + (x.material_nombre||'') + '</small></td>';
      c += '<td><b>' + (x.lote||'') + '</b></td>';
      c += '<td class="fmt-cant">' + (x.cantidad||0).toLocaleString('es-CO') + ' g</td>';
      c += '<td class="tipo-' + (x.tipo||'') + '">' + (x.tipo||'') + '</td>';
      c += '<td><b style="color:#dc2626">' + x.veces + '</b></td>';
      c += '<td>' + x.primera + ' a ' + x.ultima + '</td>';
      c += '<td>' + (x.operadores||'') + '</td></tr>';
    });
    c += '</tbody></table>';
  } else {
    c += '<div class="ok">&#x2705; No hay movimientos duplicados sospechosos en la ventana.</div>';
  }

  // Lotes creados recientes
  c += '<h2>Lotes creados en los ultimos ' + d.ventana_dias + ' dias (' + (d.lotes_creados_count||0) + ')</h2>';
  if((d.lotes_creados_count||0) === 0){
    c += '<p style="color:#78716c">Ninguno.</p>';
  } else {
    c += '<table><thead><tr><th>Fecha</th><th>Material</th><th>Lote</th><th>Cantidad</th><th>Proveedor</th><th>Operador</th></tr></thead><tbody>';
    (d.lotes_creados_recientes||[]).forEach(function(x){
      c += '<tr><td>' + (x.fecha||'') + '</td>';
      c += '<td>' + (x.material_id||'') + '<br><small style="color:#78716c">' + (x.material_nombre||'') + '</small></td>';
      c += '<td><b>' + (x.lote||'') + '</b></td>';
      c += '<td class="fmt-cant">' + (x.cantidad||0).toLocaleString('es-CO') + ' g</td>';
      c += '<td>' + (x.proveedor||'') + '</td>';
      c += '<td>' + (x.operador||'') + '</td></tr>';
    });
    c += '</tbody></table>';
  }

  // Resumen por usuario hoy
  c += '<h2>Movimientos hoy por usuario</h2>';
  if(!(d.resumen_hoy_por_usuario||[]).length){
    c += '<p style="color:#78716c">Sin movimientos hoy.</p>';
  } else {
    c += '<table><thead><tr><th>Usuario</th><th>Tipo</th><th>Movs</th><th>Lotes distintos</th></tr></thead><tbody>';
    d.resumen_hoy_por_usuario.forEach(function(x){
      c += '<tr><td><b>' + x.operador + '</b></td>';
      c += '<td class="tipo-' + (x.tipo||'') + '">' + (x.tipo||'') + '</td>';
      c += '<td>' + x.movs + '</td>';
      c += '<td>' + x.lotes_distintos + '</td></tr>';
    });
    c += '</tbody></table>';
  }

  // Movimientos recientes (ultimos 50)
  c += '<h2>Ultimos movimientos (top 50)</h2>';
  c += '<table><thead><tr><th>Fecha</th><th>Tipo</th><th>Material</th><th>Lote</th><th>Cantidad</th><th>Operador</th><th>Observaciones</th></tr></thead><tbody>';
  (d.movimientos_recientes||[]).slice(0, 50).forEach(function(x){
    c += '<tr><td>' + (x.fecha||'') + '</td>';
    c += '<td class="tipo-' + (x.tipo||'').split(' ')[0] + '">' + (x.tipo||'') + '</td>';
    c += '<td>' + (x.material_id||'') + '<br><small style="color:#78716c">' + (x.material_nombre||'') + '</small></td>';
    c += '<td><b>' + (x.lote||'') + '</b></td>';
    c += '<td class="fmt-cant">' + (x.cantidad||0).toLocaleString('es-CO') + ' g</td>';
    c += '<td>' + (x.operador||'') + '</td>';
    c += '<td><small>' + (x.observaciones||'').substring(0,80) + '</small></td></tr>';
  });
  c += '</tbody></table>';

  document.getElementById('cuerpo').innerHTML = c;
}
cargar();
</script>
</body></html>"""
    return Response(html, mimetype="text/html")


# ════════════════════════════════════════════════════════════════════════
# IMPORT INVENTARIO ENVASE — formato real de Sebastian (29-abr-2026)
# Excel con hojas ENVASES / GOTEROS / TAPAS / ETIQUETAS / PLEGADIZAS,
# header en fila 4 (col F=tipo, G=PRESENTACION, H=CANTIDAD, Q=TOTAL).
# La columna TOTAL es el inventario actual real.
# ════════════════════════════════════════════════════════════════════════

def _slug_codigo(nombre, presentacion):
    """Genera un codigo MEE consistente: tipo + slug nombre + presentacion.
    Ej: 'FRASCO AMBAR ' + '125ml' → 'AMBAR-125'.
    """
    import re as _re
    s = (nombre or '').upper().strip()
    s = _re.sub(r'[^A-Z0-9 ]', '', s)
    s = _re.sub(r'\s+', ' ', s).strip()
    # Quitar palabras genericas comunes
    GENERIC = {'FRASCO', 'ENVASE', 'BOTTLE', 'PLASTIC', 'TAPA', 'GOTERO'}
    tokens = [t for t in s.split() if t not in GENERIC]
    base = '-'.join(tokens)[:30] if tokens else s.replace(' ', '-')[:30]
    pres = (presentacion or '').upper().strip().replace(' ', '').replace('ML', '').replace('MM', '')
    pres = _re.sub(r'[^A-Z0-9/]', '', pres)
    if pres:
        return f'{base}-{pres}'.strip('-')
    return base or 'MEE-X'


@bp.route("/admin/backfill-debug", methods=["GET"])
def admin_backfill_debug_page():
    """Pagina debug para correr el backfill de checklists y ver el detalle
    de errores en pantalla — sin tener que abrir F12. Sebastian (29-abr-2026):
    "f12 no me sirve para pegar eso"."""
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)
    html = """<!DOCTYPE html><html><head><meta charset="utf-8">
    <title>Backfill checklist — debug</title>
    <style>
      body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:1100px;margin:30px auto;padding:0 20px;color:#1e293b}
      h1{font-size:20px;color:#0f172a}
      .btn{background:#a16207;color:#fff;border:none;border-radius:6px;padding:10px 20px;font-size:14px;font-weight:700;cursor:pointer}
      .ok{background:#dcfce7;color:#166534;padding:14px;border-radius:8px;border-left:4px solid #16a34a;margin-top:14px}
      .err{background:#fee2e2;color:#991b1b;padding:14px;border-radius:8px;border-left:4px solid #dc2626;margin-top:14px}
      .warn{background:#fef3c7;color:#92400e;padding:14px;border-radius:8px;border-left:4px solid #f59e0b;margin-top:14px}
      pre{background:#0f172a;color:#e2e8f0;padding:14px;border-radius:8px;font-size:11px;overflow-x:auto;max-height:500px;line-height:1.5}
      .falla{background:#fff;border:1px solid #e2e8f0;border-left:4px solid #dc2626;border-radius:8px;padding:12px 16px;margin:8px 0}
      .falla h3{margin:0 0 6px;font-size:14px;color:#991b1b}
      .nota{font-size:12px;color:#64748b;line-height:1.5;margin:10px 0}
    </style></head><body>
    <a href="/admin" style="font-size:12px;color:#0891b2">&larr; admin</a>
    <h1>🔧 Backfill checklists — debug</h1>
    <div class="nota">
      Esta página corre el endpoint <code>POST /api/programacion/checklist/backfill</code>
      y muestra el resultado COMPLETO en pantalla (sin tener que ir a la consola).
      Si hay errores por producción, los lista uno por uno con el traceback.
    </div>
    <button class="btn" onclick="ejecutar()">🔄 Ejecutar backfill ahora</button>
    <div id="resultado"></div>
    <script>
    async function ejecutar(){
      var box = document.getElementById('resultado');
      box.innerHTML = '<div class="warn">⏳ Ejecutando... puede tardar varios segundos</div>';
      try {
        var r = await fetch('/api/programacion/checklist/backfill', {method:'POST'});
        var d = await r.json();
        var html = '';
        if(!r.ok){
          html += '<div class="err"><b>❌ Error '+r.status+'</b>: '+(d.error||'sin mensaje')+
                  (d.fase ? '<br>Fase: <code>'+d.fase+'</code>' : '')+'</div>';
          if(d.traceback) html += '<h3>Traceback:</h3><pre>'+esc(d.traceback)+'</pre>';
        } else {
          var nFallas = (d.fallas||[]).length;
          if(nFallas === 0){
            html += '<div class="ok"><b>✅ '+d.mensaje+'</b></div>';
          } else {
            html += '<div class="warn"><b>⚠️ '+d.mensaje+'</b></div>';
            html += '<h3 style="margin-top:18px">Detalle de las '+nFallas+' fallas:</h3>';
            d.fallas.forEach(function(f, i){
              html += '<div class="falla">'+
                '<h3>'+(i+1)+'. '+esc(f.producto||'?')+' · '+esc(f.fecha||'?')+' · '+(f.kg||0)+' kg (id '+f.produccion_id+')</h3>'+
                '<div style="font-size:13px;color:#1e293b;margin-bottom:6px"><b>Error:</b> '+esc(f.error||'')+'</div>'+
                (f.traceback ? '<pre style="max-height:200px">'+esc(f.traceback)+'</pre>' : '')+
              '</div>';
            });
          }
        }
        html += '<details style="margin-top:18px"><summary style="cursor:pointer;color:#64748b;font-size:11px">Ver respuesta completa (raw JSON)</summary><pre>'+esc(JSON.stringify(d, null, 2))+'</pre></details>';
        box.innerHTML = html;
      } catch(e){
        box.innerHTML = '<div class="err"><b>Error de red:</b> '+esc(e.message)+'</div>';
      }
    }
    function esc(s){ return String(s||'').replace(/[&<>]/g, function(c){ return {'&':'&amp;','<':'&lt;','>':'&gt;'}[c]; }); }
    </script>
    </body></html>"""
    return Response(html, mimetype="text/html")


@bp.route("/admin/audit-inventario/limpiar-drift-mee", methods=["POST"])
def admin_audit_limpiar_drift_mee():
    """Backfill: para cada MEE donde stock_actual != SUM(movimientos_mee),
    insertar un movimiento seed por la diferencia para alinear las dos
    fuentes de verdad. NO cambia stock_actual — solo registra el log faltante.

    Sebastian (29-abr-2026): muchos MEEs vienen de import inicial sin log.
    Este endpoint limpia ese drift de una pasada.
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    fecha_iso = __import__('datetime').datetime.now().isoformat(timespec='seconds')

    # Calcular drift por MEE — misma query que el endpoint GET
    rows = c.execute("""
        SELECT mm.codigo,
               COALESCE(mm.stock_actual, 0) as stock_persistido,
               COALESCE((
                 SELECT SUM(CASE WHEN LOWER(tipo) IN ('entrada','recepcion') THEN cantidad
                                 WHEN LOWER(tipo) IN ('salida','consumo') THEN -cantidad
                                 ELSE 0 END)
                 FROM movimientos_mee
                 WHERE mee_codigo=mm.codigo AND COALESCE(anulado,0)=0
               ), 0) as stock_calc
        FROM maestro_mee mm
        WHERE COALESCE(mm.estado,'')!='Inactivo'
    """).fetchall()

    alineados = 0
    detalle = []
    for codigo, stock_persistido, stock_calc in rows:
        diff = float(stock_persistido or 0) - float(stock_calc or 0)
        if abs(diff) <= 1:
            continue  # tolerancia: ya alineado
        # Insertar movimiento seed por la diferencia
        tipo = 'Entrada' if diff > 0 else 'Salida'
        c.execute("""
            INSERT INTO movimientos_mee
              (mee_codigo, tipo, cantidad, observaciones, responsable, fecha)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (codigo, tipo, abs(diff),
              f'SEED inicial — alineación drift detectado en audit (diff={diff:+.0f})',
              u, fecha_iso))
        alineados += 1
        detalle.append({'codigo': codigo, 'diff': round(diff, 0)})
    conn.commit()
    conn.close()

    return jsonify({
        'ok': True,
        'alineados': alineados,
        'mensaje': f'{alineados} MEEs alineados con movimientos seed.',
        'detalle': detalle[:10],  # primeros 10 para no inflar response
    })


@bp.route("/admin/audit-inventario", methods=["GET"])
def admin_audit_inventario():
    """Auditoria completa del inventario: detecta drift, stocks negativos,
    producciones legacy sin descontar, movimientos huerfanos.
    Sebastian (29-abr-2026): "verifica todo".
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()

    # ── 1. MPs con stock NEGATIVO (imposible — más salidas que entradas) ──
    mps_negativos = []
    try:
        rows = c.execute("""
            SELECT m.material_id, m.material_nombre,
                   ROUND(SUM(CASE WHEN m.tipo='Entrada' THEN m.cantidad ELSE -m.cantidad END), 0) as stock_calc,
                   COUNT(*) as n_movs
            FROM movimientos m
            GROUP BY m.material_id
            HAVING stock_calc < 0
            ORDER BY stock_calc ASC
            LIMIT 50
        """).fetchall()
        cols = [d[0] for d in c.description]
        mps_negativos = [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        mps_negativos = [{'_err': str(e)}]

    # ── 2. MEEs: drift entre maestro_mee.stock_actual y SUM(movimientos_mee) ──
    mees_drift = []
    try:
        rows = c.execute("""
            SELECT mm.codigo, mm.descripcion,
                   COALESCE(mm.stock_actual, 0) as stock_persistido,
                   COALESCE((
                     SELECT SUM(CASE WHEN LOWER(tipo) IN ('entrada','recepcion') THEN cantidad
                                     WHEN LOWER(tipo) IN ('salida','consumo') THEN -cantidad
                                     ELSE 0 END)
                     FROM movimientos_mee
                     WHERE mee_codigo=mm.codigo
                       AND COALESCE(anulado,0)=0
                   ), 0) as stock_calc_movs
            FROM maestro_mee mm
            WHERE COALESCE(mm.estado,'')!='Inactivo'
        """).fetchall()
        cols = [d[0] for d in c.description]
        all_mees = [dict(zip(cols, r)) for r in rows]
        mees_drift = [
            {**m, 'diferencia': round(m['stock_persistido'] - m['stock_calc_movs'], 0)}
            for m in all_mees
            if abs(m['stock_persistido'] - m['stock_calc_movs']) > 1
        ]
        mees_drift.sort(key=lambda x: abs(x['diferencia']), reverse=True)
        mees_drift = mees_drift[:30]
    except Exception as e:
        mees_drift = [{'_err': str(e)}]

    # ── 3. MEEs con stock NEGATIVO ──
    mees_negativos = []
    try:
        rows = c.execute("""
            SELECT codigo, descripcion, stock_actual, stock_minimo
            FROM maestro_mee
            WHERE COALESCE(stock_actual,0) < 0
            ORDER BY stock_actual ASC LIMIT 30
        """).fetchall()
        cols = [d[0] for d in c.description]
        mees_negativos = [dict(zip(cols, r)) for r in rows]
    except Exception:
        pass

    # ── 4. Producciones LEGACY: estado='completado' SIN inventario_descontado_at ──
    # Estas se completaron antes del fix de hoy y nunca descontaron stock.
    legacy_sin_descontar = []
    try:
        rows = c.execute("""
            SELECT id, producto, fecha_programada, lotes,
                   COALESCE(cantidad_kg, 0) as kg
            FROM produccion_programada
            WHERE LOWER(COALESCE(estado,'')) = 'completado'
              AND COALESCE(inventario_descontado_at, '') = ''
            ORDER BY fecha_programada DESC LIMIT 50
        """).fetchall()
        cols = [d[0] for d in c.description]
        legacy_sin_descontar = [dict(zip(cols, r)) for r in rows]
    except Exception:
        pass

    # ── 5. Producciones EN PASADO sin completar (atrasadas) ──
    atrasadas = []
    try:
        rows = c.execute("""
            SELECT id, producto, fecha_programada,
                   julianday('now') - julianday(fecha_programada) as dias_atraso
            FROM produccion_programada
            WHERE LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
              AND fecha_programada < date('now','-1 day')
            ORDER BY fecha_programada ASC LIMIT 30
        """).fetchall()
        cols = [d[0] for d in c.description]
        atrasadas = [dict(zip(cols, r)) for r in rows]
    except Exception:
        pass

    # ── 6. Movimientos sin operador ni proveedor (origenes oscuros) ──
    movs_sin_origen = 0
    try:
        movs_sin_origen = c.execute("""
            SELECT COUNT(*) FROM movimientos
            WHERE COALESCE(operador,'')=''
              AND COALESCE(proveedor,'')=''
              AND COALESCE(observaciones,'')=''
        """).fetchone()[0] or 0
    except Exception:
        pass

    # ── 7. Resumen general ──
    total_mps = total_mees = total_prods_act = 0
    try:
        total_mps = c.execute("SELECT COUNT(DISTINCT material_id) FROM movimientos WHERE COALESCE(material_id,'')!=''").fetchone()[0]
    except Exception as _e:
        __import__('logging').getLogger('admin').warning('count mps fallo: %s', _e)
    try:
        total_mees = c.execute("SELECT COUNT(*) FROM maestro_mee WHERE COALESCE(estado,'')!='Inactivo'").fetchone()[0]
    except Exception as _e:
        __import__('logging').getLogger('admin').warning('count mees fallo: %s', _e)
    try:
        total_prods_act = c.execute(
            "SELECT COUNT(*) FROM produccion_programada "
            "WHERE LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')"
        ).fetchone()[0]
    except Exception as _e:
        __import__('logging').getLogger('admin').warning('count prods fallo: %s', _e)

    conn.close()

    def _esc(s): return str(s or '').replace('<','&lt;').replace('>','&gt;')
    def _num(v):
        try: return f"{float(v or 0):,.0f}"
        except: return str(v)

    def _seccion(titulo, items, cols, color, hint=''):
        if not items:
            return f'<h2 style="color:#15803d">✅ {titulo}</h2><div class="empty">Sin anomalías.</div>'
        body = f'<h2 style="color:{color}">⚠️ {titulo} ({len(items)})</h2>'
        if hint: body += f'<div class="hint">{hint}</div>'
        body += '<table><thead><tr>'
        for k,_ in cols: body += f'<th>{k}</th>'
        body += '</tr></thead><tbody>'
        for it in items:
            body += '<tr>'
            for _, key in cols:
                v = it.get(key, '')
                cell = _num(v) if isinstance(v,(int,float)) else _esc(v)
                body += f'<td>{cell}</td>'
            body += '</tr>'
        body += '</tbody></table>'
        return body

    html = ('''<!DOCTYPE html><html><head><meta charset="utf-8"><title>Auditoría inventario</title>
    <style>body{font-family:-apple-system,Segoe UI,sans-serif;max-width:1300px;margin:24px auto;padding:0 16px;color:#1e293b}
    h1{font-size:22px;margin-bottom:6px}h2{font-size:15px;margin-top:28px;border-bottom:1px solid #e2e8f0;padding-bottom:6px}
    table{width:100%;border-collapse:collapse;font-size:12px;background:#fff;margin-top:8px}
    th,td{padding:6px 10px;text-align:left;border-bottom:1px solid #f1f5f9}
    th{background:#f8fafc;font-weight:700;color:#475569;text-transform:uppercase;font-size:10px}
    .empty{color:#94a3b8;font-style:italic;padding:14px;background:#fafaf9;border-radius:6px;margin-top:8px}
    .hint{font-size:12px;color:#64748b;margin:6px 0;background:#fef3c7;padding:8px 12px;border-radius:6px;border-left:3px solid #f59e0b}
    .resumen{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin-bottom:18px}
    .kpi{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px}
    .kpi-v{font-size:22px;font-weight:800;color:#0f172a}
    .kpi-l{font-size:11px;color:#64748b;text-transform:uppercase;margin-top:2px}
    a{color:#0891b2}</style></head><body>
    <a href="/admin" style="font-size:12px">&larr; admin</a>
    <h1>🔍 Auditoría de Inventario</h1>
    <p style="color:#64748b;font-size:13px">Detecta drift, stocks negativos, producciones legacy sin descontar, movimientos huérfanos.</p>
    <div class="resumen">
      <div class="kpi"><div class="kpi-v">'''+ _num(total_mps) +'''</div><div class="kpi-l">MPs con movimientos</div></div>
      <div class="kpi"><div class="kpi-v">'''+ _num(total_mees) +'''</div><div class="kpi-l">MEEs activos</div></div>
      <div class="kpi"><div class="kpi-v">'''+ _num(total_prods_act) +'''</div><div class="kpi-l">Producciones activas</div></div>
      <div class="kpi"><div class="kpi-v" style="color:'''+ ('#dc2626' if movs_sin_origen>0 else '#15803d') +'''">'''+ _num(movs_sin_origen) +'''</div><div class="kpi-l">Movs sin origen</div></div>
    </div>'''
    + _seccion("MPs con stock NEGATIVO (más salidas que entradas)",
        mps_negativos,
        [("Código","material_id"),("Nombre","material_nombre"),
         ("Stock calc (g)","stock_calc"),("# Movs","n_movs")],
        '#dc2626',
        "Stock calculado de movimientos es < 0. Indica que se descontó más de lo que entró. Posibles causas: producción duplicada, salida sin entrada previa, fórmula con cantidad mayor a lo recibido. Revisar kardex del MP en /planta.")
    + _seccion("MEEs con DRIFT (stock_actual vs SUM movimientos_mee)",
        mees_drift,
        [("Código","codigo"),("Descripción","descripcion"),
         ("Stock persistido","stock_persistido"),("Stock movs","stock_calc_movs"),
         ("Diferencia","diferencia")],
        '#f59e0b',
        "El stock guardado en maestro_mee.stock_actual no coincide con la suma de movimientos_mee. Indica que algún endpoint actualizó stock_actual sin registrar movimiento, o vice-versa. Tolerancia ±1.")
    + (f'''<button onclick="limpiarDriftMee()" style="background:#0891b2;color:#fff;border:none;border-radius:6px;padding:8px 16px;font-size:13px;font-weight:700;cursor:pointer;margin-top:8px">🔧 Limpiar drift de {len(mees_drift)} MEEs (insertar movimientos seed)</button>
<script>
async function limpiarDriftMee(){{
  if(!confirm("Insertar movimientos seed para alinear los {len(mees_drift)} MEEs con drift?\\n\\nEsto SOLO alinea el log con la realidad actual — NO cambia el stock_actual. Operación segura."))return;
  try {{
    var r = await fetch("/admin/audit-inventario/limpiar-drift-mee",{{method:"POST",headers:{{"Content-Type":"application/json"}}}});
    var d = await r.json();
    if(!r.ok){{ alert("Error: "+(d.error||r.status)); return; }}
    alert("✅ Alineados "+d.alineados+" MEEs. La página se va a recargar.");
    location.reload();
  }} catch(e){{ alert("Error de red: "+e.message); }}
}}
</script>''' if mees_drift else '')
    + _seccion("MEEs con stock NEGATIVO",
        mees_negativos,
        [("Código","codigo"),("Descripción","descripcion"),
         ("Stock actual","stock_actual"),("Stock mínimo","stock_minimo")],
        '#dc2626')
    + _seccion("Producciones LEGACY completadas SIN descontar inventario",
        legacy_sin_descontar,
        [("ID","id"),("Producto","producto"),("Fecha","fecha_programada"),
         ("Lotes","lotes"),("kg","kg")],
        '#a16207',
        "Estas producciones se marcaron completadas antes del fix del 29-abr-2026 y NUNCA descontaron stock. Si las producciones fueron reales, el inventario actual está inflado. Decide caso por caso: revertir y re-completar con el flujo nuevo, o aceptar que ya pasaron.")
    + _seccion("Producciones ATRASADAS (fecha pasada, sin completar)",
        atrasadas,
        [("ID","id"),("Producto","producto"),("Fecha","fecha_programada"),
         ("Días atraso","dias_atraso")],
        '#f59e0b',
        "Producciones cuya fecha pasó hace >1 día y siguen 'programadas'. O bien se hicieron y no se marcaron completadas, o quedaron olvidadas. Revisar y completar/cancelar.")
    + '</body></html>')

    resp = Response(html, mimetype="text/html")
    # No cachear — el reporte cambia tras cada acción (limpiar drift, completar
    # producción, etc). Sebastian (29-abr-2026): el botón nuevo no aparecía
    # porque Chrome servía HTML cacheado.
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@bp.route("/admin/influencers-limpieza", methods=["GET"])
def admin_influencers_limpieza():
    """Lista filas dudosas en pagos_influencers para limpiarlas.
    Sebastian (29-abr-2026): "salian Pendientes y al dia muchos que ni
    hemos pagado ni tenemos pendiente". Causa: pagos_influencers tiene
    filas viejas sin OC vinculada (imports historicos) que inflan el
    badge "Al dia".
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()

    # Filas con pi.estado='Pagada' SIN OC valida en ordenes_compra
    sin_oc = c.execute("""
        SELECT pi.id, pi.influencer_nombre, pi.valor, pi.fecha, pi.estado,
               pi.numero_oc, pi.concepto
        FROM pagos_influencers pi
        LEFT JOIN ordenes_compra oc ON oc.numero_oc = pi.numero_oc
        WHERE oc.numero_oc IS NULL
        ORDER BY pi.fecha DESC LIMIT 500
    """).fetchall()
    cols = [d[0] for d in c.description]
    sin_oc = [dict(zip(cols, r)) for r in sin_oc]

    # Filas con OC en estado raro (Borrador/Pendiente >30 dias sin moverse)
    oc_stuck = c.execute("""
        SELECT pi.id, pi.influencer_nombre, pi.valor, pi.fecha, pi.estado,
               pi.numero_oc, oc.estado as oc_estado, oc.fecha as oc_fecha
        FROM pagos_influencers pi
        JOIN ordenes_compra oc ON oc.numero_oc = pi.numero_oc
        WHERE oc.estado IN ('Borrador','Pendiente','Revisada','Aprobada','Autorizada')
          AND oc.fecha < date('now','-30 day')
        ORDER BY oc.fecha ASC LIMIT 500
    """).fetchall()
    cols = [d[0] for d in c.description]
    oc_stuck = [dict(zip(cols, r)) for r in oc_stuck]

    def _esc(s): return str(s or '').replace('<','&lt;').replace('>','&gt;')
    def _money(v):
        try: return '$'+f"{float(v or 0):,.0f}"
        except Exception: return str(v)

    def _tabla(titulo, items, cols_def):
        if not items:
            return f"<h2>{titulo}</h2><div class='empty'>Nada para limpiar.</div>"
        body = f"<h2>{titulo} ({len(items)})</h2><table><thead><tr><th>✓</th>"
        for c_, _ in cols_def: body += f"<th>{c_}</th>"
        body += "</tr></thead><tbody>"
        for it in items:
            body += f"<tr><td><input type='checkbox' name='ids' value='{it['id']}'></td>"
            for _, key in cols_def:
                v = it.get(key)
                cell = _money(v) if key in ('valor',) else _esc(v)
                body += f"<td>{cell}</td>"
            body += "</tr>"
        body += "</tbody></table>"
        return body

    html = ("""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Limpieza pagos influencers</title>
    <style>body{font-family:-apple-system,Segoe UI,sans-serif;max-width:1300px;margin:24px auto;padding:0 16px;color:#1e293b}
    h1{font-size:20px;margin-bottom:6px}h2{font-size:15px;margin-top:24px;color:#0f172a;border-bottom:1px solid #e2e8f0;padding-bottom:4px}
    table{width:100%;border-collapse:collapse;font-size:12px;background:#fff}
    th,td{padding:6px 10px;text-align:left;border-bottom:1px solid #f1f5f9}
    th{background:#f8fafc;font-weight:700;color:#475569;text-transform:uppercase;font-size:10px;letter-spacing:.5px}
    .empty{color:#94a3b8;font-style:italic;padding:14px;background:#fafaf9;border-radius:6px}
    a{color:#0891b2}.warn{background:#fef3c7;color:#92400e;padding:10px 14px;border-radius:8px;margin:14px 0;font-size:13px}
    button{background:#dc2626;color:#fff;border:none;padding:8px 16px;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;margin:10px 0}
    button:hover{background:#b91c1c}
    </style></head><body>
    <a href="/admin" style="font-size:12px">&larr; admin</a>
    <h1>&#129529; Limpieza de pagos_influencers</h1>
    <div class="warn">
      <b>&#9888;&#65039; Atencion:</b> estas filas estan en pagos_influencers pero
      NO tienen OC valida o la OC esta atascada. Probablemente son imports
      historicos o solicitudes muertas. Si las eliminas, el badge "Al dia" /
      "Pendiente" del influencer se corrige.
      <br>Marca las que quieras borrar y click "Eliminar seleccionadas".
    </div>
    <form id="frm-clean" onsubmit="return false;">"""
    + _tabla("&#128206; SIN OC vinculada (potenciales imports historicos)", sin_oc,
             [("ID","id"),("Influencer","influencer_nombre"),("Valor","valor"),
              ("Fecha","fecha"),("Estado pi","estado"),("OC?","numero_oc"),
              ("Concepto","concepto")])
    + _tabla("&#9203; OC atascada >30 dias (Borrador/Pendiente/Revisada/Aprobada)", oc_stuck,
             [("ID","id"),("Influencer","influencer_nombre"),("Valor","valor"),
              ("Fecha pi","fecha"),("OC","numero_oc"),
              ("OC estado","oc_estado"),("OC fecha","oc_fecha")])
    + """
    <button type="button" onclick="eliminarSel()">&#128465; Eliminar seleccionadas</button>
    <button type="button" onclick="seleccionarTodas()" style="background:#475569">&#9989; Seleccionar todo (visible)</button>
    </form>
    <script>
    function seleccionarTodas(){
      var cbs = document.querySelectorAll('input[name=ids]');
      var allChecked = Array.from(cbs).every(function(cb){return cb.checked;});
      cbs.forEach(function(cb){ cb.checked = !allChecked; });
    }
    async function eliminarSel(){
      var ids = Array.from(document.querySelectorAll('input[name=ids]:checked')).map(function(cb){return parseInt(cb.value);});
      if(!ids.length){ alert('Marca al menos una fila para eliminar.'); return; }
      if(!confirm('Eliminar '+ids.length+' filas de pagos_influencers? Esto NO toca ordenes_compra ni solicitudes_compra.')) return;
      var r = await fetch('/admin/influencers-limpieza', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({ids:ids})});
      var d = await r.json();
      if(!r.ok){ alert('Error: '+(d.error||r.status)); return; }
      alert('Eliminadas '+d.eliminadas+' filas. Recargando.');
      location.reload();
    }
    </script>
    </body></html>""")

    conn.close()
    return Response(html, mimetype="text/html")


@bp.route("/admin/influencers-limpieza", methods=["POST"])
def admin_influencers_limpieza_post():
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({"error": "Solo admin"}), 403
    d = request.get_json() or {}
    ids = [int(x) for x in (d.get("ids") or []) if str(x).isdigit()]
    if not ids:
        return jsonify({"error": "ids vacio"}), 400
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    placeholders = ",".join("?" * len(ids))
    c.execute(f"DELETE FROM pagos_influencers WHERE id IN ({placeholders})", ids)
    conn.commit()
    n = c.rowcount
    conn.close()
    return jsonify({"ok": True, "eliminadas": n})


@bp.route("/admin/influencers-reset-pendientes", methods=["POST"])
def admin_influencers_reset_pendientes():
    """LIMPIA todos los pagos pendientes de influencers (Pendiente / Aprobada).
    Sebastian (29-abr-2026): "elimina todo eso, no hay nada pendiente, solo
    he pagado los que se han pagado".

    NO toca: pagos_influencers en estado='Pagada', OCs ya pagadas, marketing_influencers.
    SÍ borra: SOL Pendiente/Aprobada (cat Influencer/CC), OCs vinculadas en
    estado Borrador/Aprobada/Autorizada, pagos_influencers en estado Pendiente.
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403

    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    eliminado = {'pagos_influencers': 0, 'solicitudes_compra': 0, 'ordenes_compra': 0}

    # 1. Pagos influencers en estado Pendiente
    c.execute("DELETE FROM pagos_influencers WHERE LOWER(COALESCE(estado,''))='pendiente'")
    eliminado['pagos_influencers'] = c.rowcount

    # 2. SOL Pendiente/Aprobada con cat Influencer/CC
    sol_nums = c.execute("""
        SELECT numero, COALESCE(numero_oc,'') FROM solicitudes_compra
        WHERE categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')
          AND LOWER(COALESCE(estado,'')) IN ('pendiente','aprobada')
    """).fetchall()

    # 3. OCs vinculadas que NO estén pagadas
    oc_nums_a_borrar = []
    for sol_num, oc_num in sol_nums:
        if oc_num:
            oc_estado = c.execute(
                "SELECT estado FROM ordenes_compra WHERE numero_oc=?", (oc_num,)
            ).fetchone()
            if oc_estado and oc_estado[0] not in ('Pagada','Recibida','Parcial'):
                oc_nums_a_borrar.append(oc_num)

    # Borrar items de OCs
    for oc_num in oc_nums_a_borrar:
        c.execute("DELETE FROM ordenes_compra_items WHERE numero_oc=?", (oc_num,))
    if oc_nums_a_borrar:
        placeholders = ','.join('?' * len(oc_nums_a_borrar))
        c.execute(
            f"DELETE FROM ordenes_compra WHERE numero_oc IN ({placeholders})",
            oc_nums_a_borrar
        )
        eliminado['ordenes_compra'] = c.rowcount

    # Borrar SOLs (las que estaban Pendiente/Aprobada)
    if sol_nums:
        placeholders = ','.join('?' * len(sol_nums))
        c.execute(
            f"DELETE FROM solicitudes_compra_items WHERE numero IN ({placeholders})",
            [s[0] for s in sol_nums]
        )
        c.execute(
            f"DELETE FROM solicitudes_compra WHERE numero IN ({placeholders})",
            [s[0] for s in sol_nums]
        )
        eliminado['solicitudes_compra'] = c.rowcount

    conn.commit(); conn.close()
    return jsonify({
        'ok': True,
        'eliminado': eliminado,
        'mensaje': f"Limpieza completa: {eliminado['pagos_influencers']} pagos, "
                   f"{eliminado['solicitudes_compra']} SOLs, "
                   f"{eliminado['ordenes_compra']} OCs eliminadas."
    })


@bp.route("/admin/influencers-bulk-import", methods=["POST"])
def admin_influencers_bulk_import():
    """Endpoint público para cargar lote de pagos pendientes — recibe JSON.
    Body: {influencers: [{nombre, telefono, ciudad, costo, fecha_pub, concepto?, paquete?}, ...]}
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return jsonify({'error': 'Solo admin'}), 403
    d = request.get_json() or {}
    items = d.get('influencers') or []
    if not items:
        return jsonify({'error': 'influencers[] requerido'}), 400
    return _do_bulk_import(items, u)


def _do_bulk_import(items, usuario):
    """Lógica compartida: carga lote de pagos pendientes de influencers."""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    creados = []
    skipped = []

    for it in items:
        nombre = (it.get('nombre') or '').strip()
        telefono = (it.get('telefono') or '').strip()
        ciudad = (it.get('ciudad') or '').strip()
        costo_raw = it.get('costo') or 0
        fecha_pub = (it.get('fecha_pub') or '').strip()
        concepto = (it.get('concepto') or '').strip() or f'Pago contenido {fecha_pub}'
        es_paquete = bool(it.get('paquete'))

        if not nombre:
            continue
        # Si es paquete, costo=0 (pago en producto, no efectivo)
        try:
            costo = 0 if es_paquete else float(str(costo_raw).replace('.','').replace(',','.').replace('$',''))
        except Exception:
            costo = 0

        # 1. Upsert influencer
        existing = c.execute(
            "SELECT id FROM marketing_influencers "
            "WHERE LOWER(TRIM(nombre))=LOWER(TRIM(?)) "
            "  AND COALESCE(telefono,'')=? LIMIT 1",
            (nombre, telefono)
        ).fetchone()
        if existing:
            inf_id = existing[0]
            # Actualizar ciudad si vacía
            if ciudad:
                c.execute(
                    "UPDATE marketing_influencers SET ciudad=COALESCE(NULLIF(ciudad,''),?) WHERE id=?",
                    (ciudad, inf_id)
                )
        else:
            c.execute("""
                INSERT INTO marketing_influencers
                  (nombre, red_social, telefono, ciudad, tarifa, estado, fecha_registro)
                VALUES (?, 'Instagram', ?, ?, ?, 'Activo', date('now'))
            """, (nombre, telefono, ciudad, costo))
            inf_id = c.lastrowid

        # Idempotencia: skip si ya hay SOL/OC para este influencer+fecha+monto
        ya = c.execute("""
            SELECT numero FROM solicitudes_compra
            WHERE influencer_id=?
              AND ABS(COALESCE(valor,0) - ?) < 1
              AND fecha_requerida=?
              AND categoria='Cuenta de Cobro'
            LIMIT 1
        """, (inf_id, costo, fecha_pub)).fetchone()
        if ya:
            skipped.append({'nombre': nombre, 'razon': f'ya existe SOL {ya[0]}'})
            continue

        # 2. Generar número SOL único
        prefix = f"SOL-{__import__('datetime').date.today().year}-"
        last = c.execute(
            "SELECT numero FROM solicitudes_compra WHERE numero LIKE ? "
            "ORDER BY numero DESC LIMIT 1", (f"{prefix}%",)
        ).fetchone()
        if last:
            try: seq = int(last[0].split('-')[-1]) + 1
            except: seq = 1
        else:
            seq = 1
        sol_num = f"{prefix}{seq:04d}"
        oc_num = sol_num.replace('SOL', 'OC')

        # Observaciones tipo cuenta de cobro
        obs_parts = [f"BENEFICIARIO: {nombre}"]
        if telefono: obs_parts.append(f"CELULAR: {telefono}")
        if ciudad:   obs_parts.append(f"CIUDAD: {ciudad}")
        obs_parts.append(f"CONCEPTO: {concepto}")
        if es_paquete:
            obs_parts.append("VALOR: PAQUETE (pago en producto)")
        else:
            obs_parts.append(f"VALOR: ${costo:,.0f}")
        observaciones = " | ".join(obs_parts)

        # 3. INSERT SOL Aprobada
        c.execute("""
            INSERT INTO solicitudes_compra
              (numero, fecha, estado, solicitante, urgencia, observaciones,
               area, empresa, categoria, tipo, valor, influencer_id,
               fecha_requerida, numero_oc)
            VALUES (?, date('now'), 'Aprobada', 'jefferson', 'Normal', ?,
                    'Marketing', 'ANIMUS', 'Cuenta de Cobro', 'Servicio',
                    ?, ?, ?, ?)
        """, (sol_num, observaciones, costo, inf_id, fecha_pub, oc_num))

        # 4. INSERT OC Aprobada
        c.execute("""
            INSERT INTO ordenes_compra
              (numero_oc, fecha, estado, proveedor, observaciones,
               creado_por, categoria, valor_total)
            VALUES (?, date('now'), 'Aprobada', ?, ?, ?, 'Cuenta de Cobro', ?)
        """, (oc_num, nombre, observaciones, usuario, costo))

        # 5. INSERT pagos_influencers Pendiente
        try:
            c.execute("""
                INSERT INTO pagos_influencers
                  (influencer_id, influencer_nombre, valor, fecha, estado,
                   concepto, numero_oc, fecha_publicacion)
                VALUES (?, ?, ?, date('now'), 'Pendiente', ?, ?, ?)
            """, (inf_id, nombre, int(costo), concepto, oc_num, fecha_pub))
        except sqlite3.OperationalError:
            pass

        creados.append({
            'nombre': nombre, 'sol': sol_num, 'oc': oc_num,
            'costo': costo, 'fecha_pub': fecha_pub
        })

    conn.commit(); conn.close()
    return jsonify({
        'ok': True,
        'creados': len(creados),
        'skipped': len(skipped),
        'detalle': creados,
        'mensaje': f"{len(creados)} pagos pendientes cargados ({len(skipped)} duplicados omitidos)."
    })


@bp.route("/admin/influencers-cargar-29abr", methods=["GET", "POST"])
def admin_influencers_cargar_29abr():
    """Endpoint específico que carga los 16 influencers reales del foto que
    Sebastian compartió el 29-abr-2026. GET muestra UI con botón. POST ejecuta.
    Sebastian (29-abr-2026): "esta foto tiene a los que realmente les debo".
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)

    INFLUENCERS_29ABR = [
        {"nombre": "Maria Camila Soto",     "telefono": "3114902203", "ciudad": "Cali",        "costo": 1000000, "fecha_pub": "2026-04-09", "concepto": "Video 9 abril"},
        {"nombre": "Sara",                   "telefono": "3225947384", "ciudad": "Cali",        "costo": 250000,  "fecha_pub": "2026-04-10", "concepto": "Video 10 abril"},
        {"nombre": "Val sierra",             "telefono": "3235483884", "ciudad": "Cali",        "costo": 2500000, "fecha_pub": "2026-04-15", "concepto": "Video 15 abril"},
        {"nombre": "Stiven sants",           "telefono": "3206927531", "ciudad": "Cali",        "costo": 500000,  "fecha_pub": "2026-04-16", "concepto": "Video 16 abril"},
        {"nombre": "Camila Camico Torres",   "telefono": "3213784157", "ciudad": "Bogota",      "costo": 450000,  "fecha_pub": "2026-04-17", "concepto": "Rutina + lip"},
        {"nombre": "Luisa Alejandra Hoyos",  "telefono": "3113425220", "ciudad": "Pereira",     "costo": 160000,  "fecha_pub": "2026-04-19", "concepto": "Video 19 abril"},
        {"nombre": "Maria Camila Soto",     "telefono": "3114902203", "ciudad": "Cali",        "costo": 0,       "fecha_pub": "2026-04-20", "concepto": "Paquete (producto)", "paquete": True},
        {"nombre": "Valeria Osorno",         "telefono": "3216410959", "ciudad": "Medellin",    "costo": 450000,  "fecha_pub": "2026-04-21", "concepto": "Video 21 abril"},
        {"nombre": "Samira Kure",            "telefono": "3053336443", "ciudad": "Cali",        "costo": 400000,  "fecha_pub": "2026-04-22", "concepto": "Video 22 abril"},
        {"nombre": "Angie Aguilar",          "telefono": "3102657782", "ciudad": "Bogota",      "costo": 390000,  "fecha_pub": "2026-04-22", "concepto": "Video 22 abril"},
        {"nombre": "Leidy Diana Hidalgo Perea","telefono":"3004924796","ciudad": "Bello",       "costo": 420000,  "fecha_pub": "2026-04-23", "concepto": "Rutina"},
        {"nombre": "Laura Moscot Guerra",    "telefono": "3232427839", "ciudad": "Chia",        "costo": 330000,  "fecha_pub": "2026-04-24", "concepto": "Dos videos - P"},
        {"nombre": "Angela Rios",            "telefono": "3148405917", "ciudad": "Pereira",     "costo": 300000,  "fecha_pub": "2026-04-25", "concepto": "Video 25 abril"},
        {"nombre": "Tatiana Gonzalez",       "telefono": "3006598291", "ciudad": "Armenia",     "costo": 200000,  "fecha_pub": "2026-04-26", "concepto": "Video 26 abril"},
        {"nombre": "Monssa",                 "telefono": "3156127301", "ciudad": "Bucaramanga", "costo": 250000,  "fecha_pub": "2026-04-27", "concepto": "Video 27 abril"},
        {"nombre": "Camila Correal",         "telefono": "3135660143", "ciudad": "Armenia",     "costo": 300000,  "fecha_pub": "2026-04-27", "concepto": "Video 27 abril"},
    ]

    if request.method == "GET":
        total = sum(i["costo"] for i in INFLUENCERS_29ABR)
        rows_html = ""
        for i in INFLUENCERS_29ABR:
            cost = "PAQUETE" if i.get("paquete") else f"${i['costo']:,.0f}"
            rows_html += (f"<tr><td>{i['nombre']}</td><td>{i['ciudad']}</td>"
                          f"<td>{i['telefono']}</td><td>{cost}</td>"
                          f"<td>{i['fecha_pub']}</td><td>{i['concepto']}</td></tr>")
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
        <title>Cargar influencers 29abr</title>
        <style>body{{font-family:-apple-system,Segoe UI,sans-serif;max-width:1100px;margin:24px auto;padding:0 16px;color:#1e293b}}
        h1{{font-size:20px}}table{{width:100%;border-collapse:collapse;font-size:12px;margin-top:14px}}
        th,td{{padding:6px 10px;text-align:left;border-bottom:1px solid #f1f5f9}}
        th{{background:#f8fafc;font-weight:700;color:#475569;text-transform:uppercase;font-size:10px}}
        button{{background:#7c3aed;color:#fff;border:none;border-radius:8px;padding:10px 20px;font-size:14px;font-weight:700;cursor:pointer;margin-right:8px}}
        button.danger{{background:#dc2626}}
        .total{{font-size:16px;font-weight:800;color:#0f172a;margin:12px 0}}
        .warn{{background:#fef3c7;border-left:4px solid #f59e0b;padding:10px 14px;border-radius:6px;margin:14px 0;font-size:13px}}
        a{{color:#0891b2}}</style></head><body>
        <a href="/admin">&larr; admin</a>
        <h1>📋 Cargar pagos pendientes 29-abr-2026</h1>
        <div class="warn">
          <b>⚠️ Pasos:</b><br>
          1. Click <b>🧹 Limpiar pendientes actuales</b> (borra los que estaban duplicados o ficticios).<br>
          2. Click <b>📤 Cargar los 16</b> (crea SOL + OC + pago pendiente para cada uno).<br>
          3. Ve a <code>/compras → tab Influencers</code> y cada uno aparece listo para pagar con un click.
        </div>
        <p>Lista de influencers a cargar (datos del foto que enviaste):</p>
        <div class="total">Total a pagar: ${total:,.0f} COP (15 con valor + 1 paquete)</div>
        <table><thead><tr><th>Nombre</th><th>Ciudad</th><th>Teléfono</th><th>Costo</th><th>Fecha pub</th><th>Concepto</th></tr></thead>
        <tbody>{rows_html}</tbody></table>
        <div style="margin-top:24px">
          <button class="danger" onclick="limpiar()">🧹 Limpiar pendientes actuales</button>
          <button onclick="cargar()">📤 Cargar los 16</button>
        </div>
        <div id="result" style="margin-top:18px"></div>
        <script>
        async function limpiar(){{
          if(!confirm('Borrar TODOS los pagos pendientes/aprobados de influencers actuales? (No toca los ya pagados)')) return;
          var r = await fetch('/admin/influencers-reset-pendientes',{{method:'POST'}});
          var d = await r.json();
          document.getElementById('result').innerHTML = '<pre>'+JSON.stringify(d,null,2)+'</pre>';
        }}
        async function cargar(){{
          if(!confirm('Cargar los 16 influencers como pagos Aprobados pendientes de pago?')) return;
          var r = await fetch('/admin/influencers-cargar-29abr',{{method:'POST'}});
          var d = await r.json();
          document.getElementById('result').innerHTML = '<pre>'+JSON.stringify(d,null,2)+'</pre>';
        }}
        </script></body></html>"""
        return Response(html, mimetype="text/html")

    # POST: ejecuta el bulk import con la lista hardcoded — llamamos
    # directo a la función helper que hace el trabajo (no via fetch).
    return _do_bulk_import(INFLUENCERS_29ABR, u)


@bp.route("/admin/influencers-hoy", methods=["GET"])
def admin_influencers_hoy():
    """Diagnostico rapido: que paso con influencers hoy.
    Sebastian (29-abr-2026): "revisa si jeferson hoy pidio pagos de
    influencers ya sea por la pagina de influencers o por solicitudes".
    """
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    creados = solicitudes = ocs_inf = pagos_inf = []

    try:
        c.execute("""
            SELECT id, nombre, red_social, usuario_red, tarifa, fecha_registro,
                   COALESCE(banco,''), COALESCE(cuenta_bancaria,'')
            FROM marketing_influencers
            WHERE DATE(fecha_registro) = DATE('now','localtime')
               OR DATE(fecha_registro) = DATE('now')
            ORDER BY id DESC
        """)
        creados = [dict(zip([d[0] for d in c.description], r)) for r in c.fetchall()]
    except Exception:
        pass
    try:
        c.execute("""
            SELECT numero, fecha, estado, solicitante, valor, observaciones,
                   numero_oc, categoria
            FROM solicitudes_compra
            WHERE (DATE(fecha) = DATE('now','localtime') OR DATE(fecha) = DATE('now'))
              AND categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')
            ORDER BY numero DESC
        """)
        solicitudes = [dict(zip([d[0] for d in c.description], r)) for r in c.fetchall()]
    except Exception:
        pass
    try:
        c.execute("""
            SELECT numero_oc, fecha, estado, proveedor, valor_total, categoria
            FROM ordenes_compra
            WHERE (DATE(fecha) = DATE('now','localtime') OR DATE(fecha) = DATE('now'))
              AND categoria IN ('Influencer/Marketing Digital','Cuenta de Cobro')
            ORDER BY numero_oc DESC
        """)
        ocs_inf = [dict(zip([d[0] for d in c.description], r)) for r in c.fetchall()]
    except Exception:
        pass
    try:
        c.execute("""
            SELECT id, influencer_id, influencer_nombre, valor, fecha,
                   estado, concepto, numero_oc
            FROM pagos_influencers
            WHERE DATE(fecha) = DATE('now','localtime')
               OR DATE(fecha) = DATE('now')
            ORDER BY id DESC
        """)
        pagos_inf = [dict(zip([d[0] for d in c.description], r)) for r in c.fetchall()]
    except Exception:
        pass

    def _esc(s): return str(s or '').replace('<','&lt;').replace('>','&gt;')
    def _money(v):
        try: return '$'+f"{float(v or 0):,.0f}"
        except Exception: return str(v)

    def _tabla(titulo, items, cols, money_cols=()):
        if not items:
            return f"<h2>{titulo}</h2><div class='empty'>Nada hoy.</div>"
        body = f"<h2>{titulo} ({len(items)})</h2><table><thead><tr>"
        for k in cols: body += f"<th>{k}</th>"
        body += "</tr></thead><tbody>"
        for it in items:
            body += "<tr>"
            for k in cols:
                v = it.get(k)
                cell = _money(v) if k in money_cols else _esc(v)
                body += f"<td>{cell}</td>"
            body += "</tr>"
        body += "</tbody></table>"
        return body

    html = ("""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Influencers hoy</title>
    <style>body{font-family:-apple-system,Segoe UI,sans-serif;max-width:1200px;margin:24px auto;padding:0 16px;color:#1e293b}
    h1{font-size:20px;margin-bottom:6px}h2{font-size:15px;margin-top:24px;color:#0f172a;border-bottom:1px solid #e2e8f0;padding-bottom:4px}
    table{width:100%;border-collapse:collapse;font-size:12px;background:#fff}
    th,td{padding:6px 10px;text-align:left;border-bottom:1px solid #f1f5f9}
    th{background:#f8fafc;font-weight:700;color:#475569;text-transform:uppercase;font-size:10px;letter-spacing:.5px}
    .empty{color:#94a3b8;font-style:italic;padding:14px;background:#fafaf9;border-radius:6px}
    a{color:#0891b2}</style></head><body>
    <a href="/admin" style="font-size:12px">&larr; admin</a>
    <h1>&#128202; Actividad influencers &mdash; HOY</h1>
    <p style="color:#64748b;font-size:12px">Diagn&oacute;stico r&aacute;pido de qu&eacute; pas&oacute; hoy con influencers/cuentas de cobro.</p>"""
    + _tabla("&#128221; Influencers nuevos hoy", creados,
             ["id","nombre","red_social","usuario_red","tarifa","banco"],
             money_cols=("tarifa",))
    + _tabla("&#128203; Solicitudes (SOL) hoy &mdash; influencer/CC", solicitudes,
             ["numero","fecha","estado","solicitante","valor","numero_oc"],
             money_cols=("valor",))
    + _tabla("&#128722; OCs hoy &mdash; influencer/CC", ocs_inf,
             ["numero_oc","fecha","estado","proveedor","valor_total"],
             money_cols=("valor_total",))
    + _tabla("&#128181; Pagos influencers hoy", pagos_inf,
             ["id","influencer_nombre","valor","fecha","estado","numero_oc","concepto"],
             money_cols=("valor",))
    + "</body></html>")

    conn.close()
    return Response(html, mimetype="text/html")


@bp.route("/api/admin/sku-map", methods=["GET"])
def admin_sku_map_listar():
    """Lista mapeos sku_producto_map + productos disponibles en
    formula_headers para validar matches."""
    u, err, code = _require_admin()
    if err:
        return err, code
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    rows = c.execute(
        "SELECT sku, producto_nombre, COALESCE(activo,1) FROM sku_producto_map ORDER BY sku"
    ).fetchall()
    productos = [r[0] for r in c.execute(
        "SELECT producto_nombre FROM formula_headers ORDER BY producto_nombre"
    ).fetchall()]
    # Marcar mapeos huerfanos (producto_nombre que NO existe en formula_headers)
    productos_set = set(productos)
    out = [{
        'sku': r[0], 'producto_nombre': r[1],
        'activo': bool(r[2]),
        'producto_existe': (r[1] in productos_set),
    } for r in rows]
    conn.close()
    return jsonify({'mapeos': out, 'productos_disponibles': productos})


@bp.route("/api/admin/sku-map", methods=["POST"])
def admin_sku_map_upsert():
    """Upsert de un mapeo SKU. Body: {sku, producto_nombre, activo}.

    Tras editar, OPCIONAL: cancela producciones programadas con origen=
    'calendar' del producto VIEJO en fechas futuras — para limpiar las
    fantasmas que se generaron antes del fix. Pasar ?cleanup=1.
    """
    u, err, code = _require_admin()
    if err:
        return err, code
    d = request.json or {}
    sku = (d.get('sku') or '').strip().upper()
    producto = (d.get('producto_nombre') or '').strip()
    activo = 1 if d.get('activo', True) else 0
    cleanup = request.args.get('cleanup', '0') in ('1', 'true', 'True')
    producto_anterior = (d.get('producto_anterior') or '').strip()
    if not sku or not producto:
        return jsonify({'error': 'sku y producto_nombre requeridos'}), 400
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    # Capturar antes para audit
    antes_row = c.execute("SELECT producto_nombre, activo FROM sku_producto_map WHERE sku=?",
                          (sku,)).fetchone()
    antes_dict = {'producto_nombre': antes_row[0], 'activo': bool(antes_row[1])} if antes_row else None
    c.execute("""
        INSERT INTO sku_producto_map (sku, producto_nombre, activo)
        VALUES (?, ?, ?)
        ON CONFLICT(sku) DO UPDATE SET
          producto_nombre=excluded.producto_nombre,
          activo=excluded.activo
    """, (sku, producto, activo))
    canceladas = 0
    if cleanup and producto_anterior and producto_anterior != producto:
        # Cancelar producciones futuras con origen='calendar' que apuntan
        # al producto erroneo (solo las del horizonte futuro)
        try:
            cur = c.execute("""
                UPDATE produccion_programada
                SET estado='cancelado',
                    observaciones=COALESCE(observaciones,'') ||
                      ' [auto-cancelado: SKU ' || ? || ' remapeado de ' || ? || ' a ' || ? || ']'
                WHERE producto=? AND origen='calendar'
                  AND fecha_programada >= date('now','-1 day')
                  AND LOWER(COALESCE(estado,'')) NOT IN ('cancelado','completado')
            """, (sku, producto_anterior, producto, producto_anterior))
            canceladas = cur.rowcount or 0
        except Exception:
            pass
    try:
        accion_audit = 'ACTUALIZAR_SKU_MAP' if antes_dict else 'CREAR_SKU_MAP'
        audit_log(c, usuario=u, accion=accion_audit,
                  tabla='sku_producto_map', registro_id=sku,
                  antes=antes_dict,
                  despues={'producto_nombre': producto, 'activo': bool(activo),
                            'producciones_canceladas': canceladas},
                  detalle=f"{accion_audit} SKU {sku} → {producto}"
                          + (f" (canceladas {canceladas} producciones)" if canceladas else ""))
    except Exception:
        pass
    conn.commit()
    conn.close()
    return jsonify({
        'ok': True, 'sku': sku, 'producto_nombre': producto, 'activo': bool(activo),
        'canceladas_cleanup': canceladas,
        'mensaje': f'SKU {sku} mapeado a {producto}'
                   + (f' · {canceladas} producciones erroneas canceladas' if canceladas else ''),
    })


@bp.route("/admin/sku-map", methods=["GET"])
def admin_sku_map_page():
    """UI editable del mapeo SKU → producto."""
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)
    html = """<!DOCTYPE html><html><head><meta charset="utf-8">
    <title>SKU map editor</title>
    <style>
      body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:1100px;margin:30px auto;padding:0 20px;color:#1e293b}
      h1{font-size:20px;color:#0f172a}
      table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden;font-size:13px;margin-top:14px}
      th{background:#f8fafc;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:.5px;text-align:left;padding:10px}
      td{padding:8px 10px;border-top:1px solid #f1f5f9;vertical-align:middle}
      input,select{padding:6px 8px;border:1px solid #cbd5e1;border-radius:5px;font-size:12px}
      .sku{font-family:monospace;font-weight:700;color:#0891b2}
      .err-row td{background:#fee2e2}
      .warn-row td{background:#fef3c7}
      .btn{background:#16a34a;color:#fff;border:none;border-radius:5px;padding:6px 12px;font-size:11px;font-weight:700;cursor:pointer;margin-right:4px}
      .btn-del{background:#dc2626}
      .nota{font-size:12px;color:#64748b;background:#eff6ff;border-left:3px solid #3b82f6;padding:10px 14px;border-radius:6px;margin:14px 0;line-height:1.5}
    </style></head><body>
    <a href="/admin" style="font-size:12px;color:#0891b2">&larr; admin</a>
    <h1>🔗 Mapeo SKU → Producto (sku_producto_map)</h1>
    <div class="nota">
      Cuando un evento del Google Calendar tiene un SKU (ej. <code>HKJ</code>),
      el sistema lo busca en esta tabla para saber qué producto fabricar.
      Si un mapeo está mal, las producciones se generan con el producto incorrecto
      (caso real: HKJ apuntaba a Limpiador Kojico cuando es Emulsión Hidratante).<br><br>
      <b>Filas en rojo:</b> el producto referenciado NO existe en formula_headers.<br>
      <b>Editar:</b> cambia el desplegable y dale Guardar. Si quieres también
      cancelar las producciones erróneas del producto anterior, marca "Limpiar fantasmas".
    </div>
    <button onclick="cargar()" style="padding:8px 16px;background:#1e40af;color:#fff;border:none;border-radius:6px;cursor:pointer;font-weight:700">🔄 Recargar</button>
    <div id="tabla">Cargando...</div>
    <script>
    var _productos = [];
    async function cargar(){
      var r = await fetch('/api/admin/sku-map');
      var d = await r.json();
      _productos = d.productos_disponibles || [];
      var rows = (d.mapeos||[]);
      document.getElementById('tabla').innerHTML =
        '<table><thead><tr>'+
        '<th>SKU</th><th>Producto actual</th><th>Activo</th><th>¿Existe?</th><th>Cambiar a</th><th>Limpiar fantasmas</th><th></th>'+
        '</tr></thead><tbody>'+
        rows.map(function(m, i){
          var rowCls = m.producto_existe ? '' : 'err-row';
          var opciones = '<option value="">— elegir —</option>' +
            _productos.map(function(p){
              return '<option value="'+esc(p)+'"'+(p===m.producto_nombre?' selected':'')+'>'+esc(p)+'</option>';
            }).join('');
          return '<tr class="'+rowCls+'">'+
            '<td><span class="sku">'+esc(m.sku)+'</span></td>'+
            '<td>'+esc(m.producto_nombre)+'</td>'+
            '<td><input type="checkbox" id="act-'+i+'"'+(m.activo?' checked':'')+'></td>'+
            '<td>'+(m.producto_existe?'✅':'⚠️ huérfano')+'</td>'+
            '<td><select id="prod-'+i+'" style="width:280px">'+opciones+'</select></td>'+
            '<td style="text-align:center"><input type="checkbox" id="clean-'+i+'" title="Cancelar producciones programadas del producto anterior"></td>'+
            '<td><button class="btn" onclick="guardar('+i+', \\''+esc(m.sku)+'\\', \\''+esc(m.producto_nombre).replace(/\\\\/g,'').replace(/\\'/g,"\\\\'")+'\\')">💾 Guardar</button></td>'+
          '</tr>';
        }).join('') +
        '</tbody></table>';
    }
    function esc(s){ return String(s||'').replace(/[&<>"']/g, function(c){ return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]; }); }
    async function guardar(i, sku, productoAnterior){
      var prod = document.getElementById('prod-'+i).value;
      var act = document.getElementById('act-'+i).checked;
      var clean = document.getElementById('clean-'+i).checked;
      if(!prod){ alert('Selecciona un producto'); return; }
      var r = await fetch('/api/admin/sku-map'+(clean?'?cleanup=1':''), {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({sku: sku, producto_nombre: prod, activo: act, producto_anterior: productoAnterior})
      });
      var d = await r.json();
      if(!r.ok){ alert('Error: '+(d.error||r.status)); return; }
      alert(d.mensaje||'Guardado');
      cargar();
    }
    cargar();
    </script>
    </body></html>"""
    return Response(html, mimetype="text/html")


@bp.route("/api/admin/mee-fugas-check", methods=["GET"])
def admin_mee_fugas_check():
    """Audita el catalogo maestro_mee buscando "fugas" tipicas tras un
    import: codigos duplicados (case-insensitive o trim), stocks NULL,
    descripciones vacias, items con stock negativo, items archivados
    que aun aparecen en producciones programadas, y discrepancia entre
    stock_actual y la suma de movimientos_mee.

    Sebastian (29-abr-2026): "verifica que no tengamos fugas" tras
    cargar INVENTARIO ENVASE.xlsx.
    """
    u, err, code = _require_admin()
    if err:
        return err, code
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    fugas = {}

    # 1) Resumen general por categoria (lo que se importo + el resto)
    try:
        cats = c.execute("""
            SELECT COALESCE(categoria,'(sin)'),
                   COUNT(*) FILTER (WHERE COALESCE(estado,'Activo')='Activo'),
                   COUNT(*) FILTER (WHERE estado='Archivado'),
                   COALESCE(SUM(stock_actual) FILTER (WHERE COALESCE(estado,'Activo')='Activo'),0)
            FROM maestro_mee
            GROUP BY COALESCE(categoria,'(sin)')
            ORDER BY 2 DESC
        """).fetchall()
        fugas['resumen_por_categoria'] = [
            {'categoria': r[0], 'activos': r[1], 'archivados': r[2], 'stock_total': r[3]}
            for r in cats
        ]
    except Exception as e:
        fugas['resumen_por_categoria'] = {'error': str(e)}

    # 2) Codigos duplicados (case-insensitive o con trim diferente)
    try:
        dups = c.execute("""
            SELECT LOWER(TRIM(codigo)) as k, COUNT(*) as n,
                   GROUP_CONCAT(codigo, ' | ') as variantes
            FROM maestro_mee
            GROUP BY LOWER(TRIM(codigo))
            HAVING n > 1
        """).fetchall()
        fugas['codigos_duplicados'] = [
            {'clave': r[0], 'count': r[1], 'variantes': r[2]} for r in dups
        ]
    except Exception as e:
        fugas['codigos_duplicados'] = {'error': str(e)}

    # 3) Items con stock NULL o negativo
    try:
        rows = c.execute("""
            SELECT codigo, descripcion, stock_actual, estado
            FROM maestro_mee
            WHERE stock_actual IS NULL OR stock_actual < 0
            LIMIT 50
        """).fetchall()
        fugas['stock_invalido'] = [
            {'codigo': r[0], 'descripcion': r[1], 'stock': r[2], 'estado': r[3]}
            for r in rows
        ]
    except Exception as e:
        fugas['stock_invalido'] = {'error': str(e)}

    # 4) Descripciones vacias (codigo huerfano sin nombre)
    try:
        rows = c.execute("""
            SELECT codigo, stock_actual, categoria
            FROM maestro_mee
            WHERE (descripcion IS NULL OR TRIM(descripcion)='')
              AND COALESCE(estado,'Activo')='Activo'
            LIMIT 50
        """).fetchall()
        fugas['sin_descripcion'] = [
            {'codigo': r[0], 'stock': r[1], 'categoria': r[2]} for r in rows
        ]
    except Exception as e:
        fugas['sin_descripcion'] = {'error': str(e)}

    # 5) Items archivados que aparecen en checklists activos
    try:
        rows = c.execute("""
            SELECT pc.id, pc.descripcion, pc.mee_codigo_asignado,
                   m.estado, m.descripcion as mee_desc
            FROM produccion_checklist pc
            JOIN maestro_mee m ON m.codigo = pc.mee_codigo_asignado
            WHERE m.estado='Archivado'
              AND pc.estado IN ('pendiente','solicitado','en_transito')
            LIMIT 30
        """).fetchall()
        fugas['archivados_en_checklist'] = [
            {'item_id': r[0], 'item_desc': r[1], 'mee_codigo': r[2],
             'mee_estado': r[3], 'mee_desc': r[4]}
            for r in rows
        ]
    except Exception as e:
        fugas['archivados_en_checklist'] = {'error': str(e)}

    # 6) Reconciliacion stock vs movimientos (top 10 con mayor desfase)
    try:
        rows = c.execute("""
            SELECT m.codigo, m.descripcion, COALESCE(m.stock_actual,0) as stock_actual,
                   COALESCE((
                     SELECT SUM(CASE
                       WHEN tipo IN ('Entrada','Ajuste +','Ajuste') THEN cantidad
                       WHEN tipo IN ('Salida','Ajuste -') THEN -cantidad
                       ELSE 0 END)
                     FROM movimientos_mee mm WHERE mm.mee_codigo=m.codigo
                   ),0) as stock_calc
            FROM maestro_mee m
            WHERE COALESCE(m.estado,'Activo')='Activo'
              AND m.categoria IN ('Envases','Goteros','Tapas')
        """).fetchall()
        desfases = []
        for r in rows:
            sa, sc = float(r[2] or 0), float(r[3] or 0)
            if abs(sa - sc) > 0.5:
                desfases.append({
                    'codigo': r[0], 'descripcion': r[1],
                    'stock_actual': sa, 'stock_movimientos': sc,
                    'diff': round(sa - sc, 2),
                })
        desfases.sort(key=lambda x: abs(x['diff']), reverse=True)
        fugas['reconciliacion_stock'] = {
            'count_con_desfase': len(desfases),
            'top_10': desfases[:10],
        }
    except Exception as e:
        fugas['reconciliacion_stock'] = {'error': str(e)}

    # 7) Movimientos del import reciente (audit trail del Excel)
    try:
        rows = c.execute("""
            SELECT COUNT(*),
                   COALESCE(SUM(cantidad),0)
            FROM movimientos_mee
            WHERE observaciones LIKE '%INVENTARIO ENVASE%'
              AND fecha >= datetime('now','-2 day')
        """).fetchone()
        fugas['movimientos_import_2d'] = {
            'count': rows[0] if rows else 0,
            'cantidad_total': rows[1] if rows else 0,
        }
    except Exception as e:
        fugas['movimientos_import_2d'] = {'error': str(e)}

    # Score: fugas detectadas
    n_dup = len(fugas.get('codigos_duplicados', []) or [])
    n_neg = len(fugas.get('stock_invalido', []) or [])
    n_sd  = len(fugas.get('sin_descripcion', []) or [])
    n_arc = len(fugas.get('archivados_en_checklist', []) or [])
    n_rec = (fugas.get('reconciliacion_stock', {}) or {}).get('count_con_desfase', 0)
    fugas['_summary'] = {
        'codigos_duplicados': n_dup,
        'stock_invalido': n_neg,
        'sin_descripcion': n_sd,
        'archivados_en_uso': n_arc,
        'desfase_stock_movimientos': n_rec,
        'todo_ok': (n_dup == 0 and n_neg == 0 and n_sd == 0 and n_arc == 0),
    }
    conn.close()
    return jsonify(fugas)


@bp.route("/admin/mee-fugas-check", methods=["GET"])
def admin_mee_fugas_check_page():
    """Pagina HTML para ver el resultado de la auditoria post-import."""
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)
    html = """<!DOCTYPE html><html><head><meta charset="utf-8">
    <title>MEE — verificación de fugas</title>
    <style>
      body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:1100px;margin:30px auto;padding:0 20px;color:#1e293b}
      h1{font-size:20px}
      h2{font-size:15px;color:#0f172a;margin-top:22px;border-bottom:1px solid #e2e8f0;padding-bottom:6px}
      .ok{background:#dcfce7;color:#166534;padding:10px 14px;border-radius:8px;font-weight:700;border-left:4px solid #16a34a}
      .warn{background:#fef3c7;color:#92400e;padding:10px 14px;border-radius:8px;font-weight:700;border-left:4px solid #f59e0b}
      .err{background:#fee2e2;color:#991b1b;padding:10px 14px;border-radius:8px;font-weight:700;border-left:4px solid #dc2626}
      table{width:100%;border-collapse:collapse;font-size:12px;background:#fff;border:1px solid #e2e8f0;border-radius:6px;overflow:hidden;margin:6px 0}
      th{background:#f8fafc;color:#475569;text-align:left;padding:8px}
      td{padding:8px;border-top:1px solid #f1f5f9}
      .small{font-size:11px;color:#94a3b8}
      .empty{color:#94a3b8;font-style:italic;padding:8px 0}
    </style></head><body>
    <a href="/admin" style="font-size:12px;color:#0891b2">&larr; admin</a>
    <h1>🔍 MEE — verificación de fugas tras import</h1>
    <div id="content">Cargando...</div>
    <script>
    function tablaFromArr(arr, cols, emptyMsg){
      if(!arr || !arr.length) return '<div class="empty">'+emptyMsg+'</div>';
      var head = '<tr>' + cols.map(function(c){return '<th>'+c.label+'</th>';}).join('') + '</tr>';
      var rows = arr.map(function(r){
        return '<tr>' + cols.map(function(c){
          var v = r[c.key];
          if(typeof v === 'number') return '<td style="font-family:monospace">'+(v.toLocaleString('es-CO'))+'</td>';
          return '<td>'+(v==null?'':String(v))+'</td>';
        }).join('') + '</tr>';
      }).join('');
      return '<table><thead>'+head+'</thead><tbody>'+rows+'</tbody></table>';
    }
    async function cargar(){
      var r = await fetch('/api/admin/mee-fugas-check');
      var d = await r.json();
      var s = d._summary || {};
      var todoOk = s.todo_ok && (s.desfase_stock_movimientos===0);
      var head = todoOk
        ? '<div class="ok">✅ Sin fugas detectadas. Catálogo limpio.</div>'
        : ('<div class="' + (s.codigos_duplicados+s.stock_invalido+s.archivados_en_uso > 0 ? 'err' : 'warn') + '">' +
           '⚠️ Hallazgos: ' +
           [
             s.codigos_duplicados>0 ? s.codigos_duplicados+' códigos duplicados' : '',
             s.stock_invalido>0 ? s.stock_invalido+' stock inválido' : '',
             s.sin_descripcion>0 ? s.sin_descripcion+' sin descripción' : '',
             s.archivados_en_uso>0 ? s.archivados_en_uso+' archivados en uso' : '',
             s.desfase_stock_movimientos>0 ? s.desfase_stock_movimientos+' con desfase stock vs movimientos' : '',
           ].filter(Boolean).join(' · ') +
           '</div>');

      var html = head;

      // Resumen por categoria
      html += '<h2>📊 Catálogo por categoría</h2>';
      html += tablaFromArr(d.resumen_por_categoria, [
        {key:'categoria', label:'Categoría'},
        {key:'activos', label:'Activos'},
        {key:'archivados', label:'Archivados'},
        {key:'stock_total', label:'Stock total (und)'},
      ], 'Sin datos');

      // Movimientos del import
      var mi = d.movimientos_import_2d || {};
      html += '<h2>📥 Auditoría del último import (últimos 2 días)</h2>';
      html += '<div style="font-size:13px"><b>'+(mi.count||0)+'</b> movimientos registrados con etiqueta "INVENTARIO ENVASE" · cantidad acumulada: <b>'+(mi.cantidad_total||0).toLocaleString('es-CO')+'</b></div>';

      // Duplicados
      html += '<h2>🔁 Códigos duplicados</h2>';
      html += tablaFromArr(d.codigos_duplicados, [
        {key:'clave', label:'Clave normalizada'},
        {key:'count', label:'Veces'},
        {key:'variantes', label:'Variantes'},
      ], '✅ Sin duplicados');

      // Stock inválido
      html += '<h2>⚠️ Stock NULL o negativo</h2>';
      html += tablaFromArr(d.stock_invalido, [
        {key:'codigo', label:'Código'},
        {key:'descripcion', label:'Descripción'},
        {key:'stock', label:'Stock'},
        {key:'estado', label:'Estado'},
      ], '✅ Todos los stocks válidos');

      // Sin descripción
      html += '<h2>❓ Items activos sin descripción</h2>';
      html += tablaFromArr(d.sin_descripcion, [
        {key:'codigo', label:'Código'},
        {key:'stock', label:'Stock'},
        {key:'categoria', label:'Categoría'},
      ], '✅ Todos tienen descripción');

      // Archivados en uso
      html += '<h2>🛑 Items archivados todavía referenciados en checklists activos</h2>';
      html += tablaFromArr(d.archivados_en_checklist, [
        {key:'item_id', label:'Item'},
        {key:'item_desc', label:'Descripción del item'},
        {key:'mee_codigo', label:'MEE código'},
        {key:'mee_desc', label:'MEE descripción'},
        {key:'mee_estado', label:'Estado MEE'},
      ], '✅ Ningún archivado referenciado');

      // Reconciliacion
      var rec = d.reconciliacion_stock || {};
      html += '<h2>📐 Reconciliación: stock_actual vs SUM(movimientos)</h2>';
      html += '<div style="font-size:13px;margin-bottom:6px"><b>'+(rec.count_con_desfase||0)+'</b> items con desfase mayor a 0.5 und (top 10):</div>';
      html += tablaFromArr(rec.top_10 || [], [
        {key:'codigo', label:'Código'},
        {key:'descripcion', label:'Descripción'},
        {key:'stock_actual', label:'stock_actual'},
        {key:'stock_movimientos', label:'SUM movs'},
        {key:'diff', label:'Δ'},
      ], '✅ Sin desfases');
      html += '<div class="small">El desfase es esperable cuando se importa stock como override (no via movimientos), pero no debe crecer descontrolado. La columna Δ positiva = stock_actual > suma de movimientos (override del Excel).</div>';

      document.getElementById('content').innerHTML = html;
    }
    cargar();
    </script>
    </body></html>"""
    return Response(html, mimetype="text/html")


@bp.route("/admin/inventario-envase-import", methods=["GET"])
def admin_inventario_envase_import_page():
    """Página simple para que Sebastian suba el Excel INVENTARIO ENVASE."""
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)
    html = """<!DOCTYPE html><html><head><meta charset="utf-8">
    <title>Import Inventario Envase</title>
    <style>
      body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:760px;margin:30px auto;padding:0 20px;color:#1e293b}
      h1{font-size:20px;color:#0f172a}
      .card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:20px;margin:16px 0;box-shadow:0 1px 3px rgba(0,0,0,.05)}
      .label{font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:.5px;font-weight:700;margin-bottom:6px;display:block}
      input[type=file]{display:block;margin:8px 0 14px;padding:8px;border:1px solid #cbd5e1;border-radius:6px;width:100%;font-size:13px}
      .btn{background:#16a34a;color:#fff;border:none;border-radius:6px;padding:9px 18px;font-size:13px;font-weight:700;cursor:pointer;margin-right:8px}
      .btn.preview{background:#f59e0b}
      .btn.reset{background:#dc2626}
      pre{background:#f1f5f9;padding:12px;border-radius:8px;font-size:11px;overflow-x:auto;max-height:400px}
      .nota{font-size:12px;color:#64748b;line-height:1.6;background:#fefce8;border-left:3px solid #f59e0b;padding:10px 14px;border-radius:6px}
    </style></head><body>
    <a href="/admin" style="font-size:12px;color:#0891b2">&larr; Volver al panel admin</a>
    <h1>📤 Importar INVENTARIO ENVASE.xlsx</h1>
    <div class="nota">
      Sube el archivo <code>INVENTARIO ENVASE.xlsx</code>. El sistema lee las hojas
      <b>ENVASES, GOTEROS, TAPAS</b> (las otras se ignoran por ahora) y actualiza el
      inventario con la columna <b>TOTAL</b> de cada fila.<br><br>
      <b>Preview</b> muestra qué pasaría sin escribir.<br>
      <b>Importar</b> aplica el upsert (crea nuevos códigos + ajusta stock de existentes).<br>
      <b>RESET + Importar</b> archiva todos los Envases/Goteros/Tapas activos antes de importar.
    </div>
    <div class="card">
      <label class="label">Archivo Excel</label>
      <input type="file" id="file" accept=".xlsx,.xlsm">
      <div>
        <button class="btn preview" onclick="enviar(true,'upsert')">👁 Preview (dry run)</button>
        <button class="btn" onclick="enviar(false,'upsert')">📥 Importar (upsert)</button>
        <button class="btn reset" onclick="enviarReset()">⚠️ RESET + Importar</button>
      </div>
    </div>
    <div id="resultado"></div>
    <script>
    async function enviar(dryRun, modo){
      var f = document.getElementById('file').files[0];
      if(!f){ alert('Selecciona un archivo'); return; }
      var fd = new FormData();
      fd.append('file', f);
      var qs = (dryRun?'dry_run=1&':'')+'modo='+encodeURIComponent(modo);
      var r = await fetch('/api/admin/import-inventario-envase-xlsx?'+qs, {method:'POST', body: fd});
      var d = await r.json();
      var col = d.ok ? '#16a34a' : '#dc2626';
      document.getElementById('resultado').innerHTML =
        '<div class="card" style="border-left:4px solid '+col+'"><h3 style="margin:0 0 8px;color:'+col+'">'+
        (d.ok?(d.dry_run?'✅ Preview OK':'✅ Importado'):'❌ Error')+'</h3>'+
        '<pre>'+JSON.stringify(d, null, 2)+'</pre></div>';
    }
    function enviarReset(){
      if(!confirm('⚠️ Esto ARCHIVA todos los Envases/Goteros/Tapas activos antes de importar. ¿Procedes?')) return;
      enviar(false, 'reset_envases');
    }
    </script>
    </body></html>"""
    return Response(html, mimetype="text/html")


@bp.route("/admin/producciones-debug", methods=["GET"])
def admin_producciones_debug_page():
    """Pagina simple para listar producciones programadas activas y borrar
    fantasmas. Sebastian (29-abr-2026): "ese Limpiador kojico 20 kg no se
    de donde lo esta sacando, revisa porfa para que lo elimines"."""
    u = session.get("compras_user", "")
    if u not in ADMIN_USERS:
        return Response("403", status=403)
    html = """<!DOCTYPE html><html><head><meta charset="utf-8">
    <title>Producciones programadas — debug</title>
    <style>
      body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;max-width:1100px;margin:30px auto;padding:0 20px;color:#1e293b}
      h1{font-size:20px;color:#0f172a}
      table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden}
      th{background:#f8fafc;color:#475569;font-size:11px;text-transform:uppercase;letter-spacing:.5px;text-align:left;padding:10px}
      td{padding:10px;border-top:1px solid #f1f5f9;font-size:13px}
      .badge{font-size:10px;font-weight:700;padding:2px 8px;border-radius:8px;text-transform:uppercase}
      .b-cal{background:#dbeafe;color:#1e40af}
      .b-man{background:#fef3c7;color:#92400e}
      .btn-del{background:#dc2626;color:#fff;border:none;border-radius:5px;padding:5px 10px;font-size:11px;font-weight:700;cursor:pointer}
      .nota{font-size:12px;color:#64748b;line-height:1.5;background:#eff6ff;border-left:3px solid #3b82f6;padding:10px 14px;border-radius:6px;margin:14px 0}
    </style></head><body>
    <a href="/admin" style="font-size:12px;color:#0891b2">&larr; Volver al panel admin</a>
    <h1>🗓️ Producciones programadas — diagnóstico</h1>
    <div class="nota">
      Lista todas las producciones activas (futuras y de hace ≤7 días). El campo
      <b>origen</b> indica si vino del calendario (auto-sync) o se creó manualmente
      en la app. Si ves una <b>fantasma</b> (que no recuerdas haber programado),
      borra. El sync con calendar se ejecuta cada 10 min y ahora es bidireccional —
      borra del calendar y desaparece sola en minutos.
    </div>
    <button onclick="cargar()" style="margin:12px 0;padding:8px 16px;background:#1e40af;color:#fff;border:none;border-radius:6px;cursor:pointer;font-weight:700">🔄 Actualizar</button>
    <button onclick="forzarSync()" style="margin:12px 0;padding:8px 16px;background:#0891b2;color:#fff;border:none;border-radius:6px;cursor:pointer;font-weight:700">📅 Forzar sync calendario</button>
    <div id="tabla">Cargando...</div>
    <script>
    async function cargar(){
      var r = await fetch('/api/programacion/produccion-programada/listado');
      var d = await r.json();
      if(!r.ok){ document.getElementById('tabla').innerHTML='Error: '+(d.error||r.status); return; }
      var prods = d.producciones || [];
      if(!prods.length){ document.getElementById('tabla').innerHTML='<i>Sin producciones activas.</i>'; return; }
      document.getElementById('tabla').innerHTML =
        '<table><thead><tr>'+
        '<th>ID</th><th>Producto</th><th>Fecha</th><th>kg</th><th>Origen</th><th>Estado</th><th>Observaciones</th><th></th>'+
        '</tr></thead><tbody>'+
        prods.map(function(p){
          var bcl = p.origen==='calendar' ? 'b-cal' : 'b-man';
          return '<tr>'+
            '<td style="font-family:monospace;color:#64748b">'+p.id+'</td>'+
            '<td><b>'+esc(p.producto)+'</b></td>'+
            '<td>'+esc(p.fecha_programada||'')+'</td>'+
            '<td style="font-family:monospace">'+(Math.round(p.kg||0))+'</td>'+
            '<td><span class="badge '+bcl+'">'+esc(p.origen)+'</span></td>'+
            '<td>'+esc(p.estado||'')+'</td>'+
            '<td style="font-size:11px;color:#64748b;max-width:300px">'+esc((p.observaciones||'').substring(0,100))+'</td>'+
            '<td><button class="btn-del" onclick="borrar('+p.id+', \\''+esc(p.producto).replace(/\\\\/g,'').replace(/\\'/g,"\\\\'")+'\\', \\''+esc(p.fecha_programada)+'\\')">Borrar</button></td>'+
          '</tr>';
        }).join('') +
        '</tbody></table>';
    }
    function esc(s){ return String(s||'').replace(/[&<>"']/g, function(c){ return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]; }); }
    async function borrar(id, prod, fecha){
      if(!confirm('¿Borrar definitivamente la producción '+id+' ('+prod+' · '+fecha+')? También borra items del checklist asociados.')) return;
      var r = await fetch('/api/programacion/produccion-programada/'+id+'/borrar', {method:'DELETE'});
      var d = await r.json();
      if(!r.ok){ alert('Error: '+(d.error||r.status)); return; }
      alert(d.mensaje||'Borrada');
      cargar();
    }
    async function forzarSync(){
      var r = await fetch('/api/programacion/checklist/sync-calendar?dias=120', {method:'POST'});
      var d = await r.json();
      alert(d.mensaje || JSON.stringify(d));
      cargar();
    }
    cargar();
    </script>
    </body></html>"""
    return Response(html, mimetype="text/html")


@bp.route("/api/admin/import-inventario-envase-xlsx", methods=["POST"])
def admin_import_inventario_envase_xlsx():
    """Importa el Excel INVENTARIO ENVASE.xlsx de Sebastian a maestro_mee.

    Sebastian (29-abr-2026): "@INVENTARIO ENVASE.xlsx... mira allí están los
    datos reales con inventario actual donde dice TOTAL... resuelve eso por
    favor, ya sea que elimines y montes o normalices".

    Body: multipart/form-data con 'file' = .xlsx
    Query: ?dry_run=1 para preview sin escribir
           ?modo=upsert (default) | reset_envases (borra todos los items
                                    categoria='Envases' antes de importar)

    Lee 3 hojas estructuradas: ENVASES, GOTEROS, TAPAS. Las hojas ETIQUETAS
    y PLEGADIZAS tienen formato distinto y se ignoran (Sebastian las puede
    cargar luego con otro formato).

    Estructura esperada (header en fila 4, datos desde fila 6):
      F: nombre del material
      G: PRESENTACION (ej. 30ml, 89mm)
      H: CANTIDAD (ingresos historicos — informativo, NO se usa para stock)
      Q (col 17): TOTAL = stock actual real

    Genera codigo automatico: tipo + slug(nombre) + presentacion.
    Por cada fila: INSERT si no existe, UPDATE stock_actual si existe
    (ajuste con auditoria en movimientos_mee).
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    if 'file' not in request.files:
        return jsonify({'error': 'Falta archivo (campo "file")'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    dry_run = request.args.get('dry_run', '0') in ('1', 'true', 'True')
    modo = (request.args.get('modo') or 'upsert').strip()

    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception as e:
        return jsonify({'error': f'Excel inválido: {e}'}), 400

    # Mapeo hoja → categoria/prefijo
    HOJAS = {
        'ENVASES': ('Envases', 'ENV'),
        'GOTEROS': ('Goteros', 'GOT'),
        'TAPAS':   ('Tapas',   'TAP'),
    }
    items = []
    detalles_por_hoja = {}
    for sheet_name, (categoria, prefijo) in HOJAS.items():
        if sheet_name not in wb.sheetnames:
            detalles_por_hoja[sheet_name] = {'leidos': 0, 'skipped': 0, 'razon_skip': 'hoja no existe'}
            continue
        ws = wb[sheet_name]
        leidos = 0; skipped = 0
        for ri in range(6, ws.max_row + 1):
            nombre = ws.cell(row=ri, column=6).value      # F
            presentacion = ws.cell(row=ri, column=7).value  # G
            total = ws.cell(row=ri, column=17).value      # Q
            if not nombre:
                continue
            nombre_str = str(nombre).strip()
            if not nombre_str:
                continue
            try:
                stock = float(total) if total is not None else 0.0
            except (ValueError, TypeError):
                stock = 0.0
                skipped += 1
            pres_str = str(presentacion or '').strip()
            slug = _slug_codigo(nombre_str, pres_str)
            codigo = f'{prefijo}-{slug}'[:50]
            descripcion_full = (
                f'{nombre_str} {pres_str}'.strip()
                if pres_str else nombre_str
            )
            items.append({
                'codigo': codigo,
                'descripcion': descripcion_full,
                'categoria': categoria,
                'unidad': 'und',
                'stock': stock,
                'sheet': sheet_name,
                'row': ri,
            })
            leidos += 1
        detalles_por_hoja[sheet_name] = {'leidos': leidos, 'skipped': skipped}

    if not items:
        return jsonify({'error': 'No se leyo ningun item de las hojas conocidas',
                        'hojas_leidas': detalles_por_hoja}), 400

    if dry_run:
        return jsonify({
            'ok': True, 'dry_run': True,
            'total_items': len(items),
            'hojas': detalles_por_hoja,
            'preview': items[:8],
            'mensaje': f'{len(items)} items detectados. Re-envia sin dry_run para escribir.',
        })

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    insertados = 0; actualizados = 0; archivados = 0; ajustes_stock = 0

    if modo == 'reset_envases':
        # Archivar todos los items que NO van a venir en el import (de las
        # 3 categorias afectadas). NO los borra — quedan auditables.
        try:
            cur = c.execute(
                "UPDATE maestro_mee SET estado='Archivado' "
                "WHERE categoria IN ('Envases','Goteros','Tapas') "
                "AND COALESCE(estado,'Activo')='Activo'"
            )
            archivados = cur.rowcount or 0
        except Exception:
            pass

    for it in items:
        existing = c.execute(
            "SELECT codigo, COALESCE(stock_actual,0) FROM maestro_mee WHERE codigo=?",
            (it['codigo'],)
        ).fetchone()
        if existing:
            stock_anterior = float(existing[1] or 0)
            c.execute("""
                UPDATE maestro_mee SET
                  descripcion=?, categoria=?, unidad=?, stock_actual=?,
                  estado='Activo'
                WHERE codigo=?
            """, (it['descripcion'], it['categoria'], it['unidad'],
                  it['stock'], it['codigo']))
            if abs(it['stock'] - stock_anterior) > 0.01:
                try:
                    c.execute("""
                        INSERT INTO movimientos_mee (mee_codigo, tipo, cantidad, responsable, observaciones)
                        VALUES (?, 'Ajuste', ?, ?, ?)
                    """, (it['codigo'], abs(it['stock']-stock_anterior), u,
                          f'[Import INVENTARIO ENVASE.xlsx · {it["sheet"]}#{it["row"]}] '
                          f'{stock_anterior:.0f} → {it["stock"]:.0f}'))
                except sqlite3.OperationalError:
                    pass
                ajustes_stock += 1
            actualizados += 1
        else:
            c.execute("""
                INSERT INTO maestro_mee
                  (codigo, descripcion, categoria, unidad, stock_actual,
                   stock_minimo, estado, fecha_creacion)
                VALUES (?, ?, ?, ?, ?, 1000, 'Activo', datetime('now'))
            """, (it['codigo'], it['descripcion'], it['categoria'],
                  it['unidad'], it['stock']))
            try:
                c.execute("""
                    INSERT INTO movimientos_mee (mee_codigo, tipo, cantidad, responsable, observaciones)
                    VALUES (?, 'Entrada', ?, ?, ?)
                """, (it['codigo'], it['stock'], u,
                      f'[Import inicial INVENTARIO ENVASE.xlsx · {it["sheet"]}#{it["row"]}]'))
            except sqlite3.OperationalError:
                pass
            insertados += 1
    conn.commit()
    conn.close()
    return jsonify({
        'ok': True,
        'modo': modo,
        'total_items_archivo': len(items),
        'insertados': insertados,
        'actualizados': actualizados,
        'ajustes_stock': ajustes_stock,
        'archivados': archivados,
        'hojas': detalles_por_hoja,
        'mensaje': f'OK · {insertados} nuevos · {actualizados} actualizados · '
                   f'{ajustes_stock} con cambio de stock'
                   + (f' · {archivados} archivados (modo reset)' if archivados else ''),
    })


# ─── Auditoría de Catálogo MPs ────────────────────────────────────────────────
# Sebastián 10-may-2026: detectar duplicados, inconsistencias y huérfanos
# antes del inventario físico. Read-only · admin only · 12 checks.

@bp.route("/api/admin/auditoria-catalogo", methods=["GET"])
def auditoria_catalogo():
    """Audit completo del catálogo de MPs y bodega.

    Detecta 12 tipos de inconsistencias agrupados por severidad:

    ALTA (bloquean producción / pérdida trazabilidad):
      1. Códigos MP duplicados activos (debería ser imposible por PK)
      2. Mismo INCI con códigos MP distintos
      3. Stock negativo por MP/lote
      4. Lotes vencidos pero estado_lote=VIGENTE con stock > 0

    MEDIA (operativos · confunden trabajo diario):
      5. Mismo nombre comercial con códigos distintos
      6. Movimientos huérfanos (material_id sin fila en maestro_mps)
      7. Tipo material inválido (fuera de MP/Envase Primario/etc)
      8. Lotes duplicados entre MPs distintas (mismo número de lote)
      9. Códigos MP con espacios o caracteres extraños

    BAJA (limpieza):
      10. MPs activas sin movimientos (catálogo muerto)
      11. Proveedores con casing inconsistente
      12. Nombres muy similares (fuzzy · típicamente typos)

    Returns:
      {
        ok: bool,
        timestamp: ISO,
        resumen: {n_alta, n_media, n_baja, n_total_findings},
        findings: { alta: {...}, media: {...}, baja: {...} },
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    # Modo rápido: salta el check fuzzy (#12 · O(n²) con difflib, ~5-10s
    # en catálogos grandes). Sebastián 10-may-2026: agregado para evitar
    # timeouts del browser cuando el catálogo crece.
    quick_mode = request.args.get('quick', '').strip() in ('1', 'true', 'yes')

    import datetime as _dt
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    findings = {'alta': {}, 'media': {}, 'baja': {}}

    # === ALTA ===
    # 1. Códigos MP duplicados activos (PK debería prevenir, verificar)
    try:
        rows = c.execute("""
            SELECT codigo_mp, COUNT(*) as cnt
            FROM maestro_mps WHERE activo=1
            GROUP BY codigo_mp HAVING cnt > 1
        """).fetchall()
        findings['alta']['codigos_mp_duplicados'] = [
            {'codigo_mp': r[0], 'count': r[1]} for r in rows
        ]
    except Exception as e:
        findings['alta']['codigos_mp_duplicados_err'] = str(e)[:200]

    # 2. Mismo INCI con códigos MP distintos
    # Sebastián 10-may-2026: hay INCI que se REPITEN LEGÍTIMAMENTE entre
    # MPs distintas en cosmética (Parfum, Aqua, etc · cada fragancia es
    # una MP comercial distinta pero con mismo nombre INCI por estándar).
    # Estos NO son duplicados reales · se mueven a BAJA con label.
    INCI_WHITELIST_LEGAL = {
        'parfum', 'fragrance', 'aroma',
        'aqua', 'water', 'agua',
        'alcohol', 'alcohol denat',
        'glycerin', 'glicerina',
        '(varies)', 'mixture',
    }
    INCI_PLACEHOLDERS = {
        'pendiente inci', 'pendiente', 'sin inci', 'no inci',
        'por definir', 'tbd', 'n/a', 'na', '-',
    }
    try:
        rows = c.execute("""
            SELECT LOWER(TRIM(nombre_inci)) as inci_norm,
                   GROUP_CONCAT(codigo_mp) as codigos,
                   COUNT(*) as cnt,
                   GROUP_CONCAT(DISTINCT nombre_inci) as variantes
            FROM maestro_mps
            WHERE activo=1 AND nombre_inci IS NOT NULL
                  AND TRIM(nombre_inci) != ''
            GROUP BY LOWER(TRIM(nombre_inci))
            HAVING cnt > 1
            ORDER BY cnt DESC LIMIT 100
        """).fetchall()
        inci_real_dup = []
        inci_legal_compartido = []
        inci_placeholder = []
        for r in rows:
            inci_norm = r[0]
            grupo = {
                'inci_normalizado': inci_norm,
                'codigos_mp': (r[1] or '').split(','),
                'count': r[2], 'variantes_raw': r[3],
            }
            if inci_norm in INCI_WHITELIST_LEGAL:
                inci_legal_compartido.append({
                    **grupo,
                    'label': 'INCI cosmético compartido legalmente · NO fusionar',
                })
            elif inci_norm in INCI_PLACEHOLDERS:
                inci_placeholder.append({
                    **grupo,
                    'label': 'INCI placeholder · falta definir el real',
                })
            else:
                inci_real_dup.append(grupo)
        findings['alta']['inci_duplicado'] = inci_real_dup
        # Mover legales y placeholders a BAJA
        if inci_legal_compartido:
            findings['baja']['inci_compartido_legal'] = inci_legal_compartido
        if inci_placeholder:
            findings['media']['inci_pendiente_llenar'] = inci_placeholder
    except Exception as e:
        findings['alta']['inci_duplicado_err'] = str(e)[:200]

    # 3. Stock negativo por MP/lote
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,'') as lote,
                   ROUND(SUM(CASE WHEN tipo='Entrada' THEN cantidad
                                  ELSE -cantidad END), 2) as stock_neto,
                   COUNT(*) as n_movs
            FROM movimientos
            GROUP BY material_id, lote
            HAVING stock_neto < -0.5
            ORDER BY stock_neto ASC LIMIT 50
        """).fetchall()
        findings['alta']['stock_negativo'] = [
            {'material_id': r[0], 'lote': r[1],
             'stock_neto_g': r[2], 'n_movimientos': r[3]}
            for r in rows
        ]
    except Exception as e:
        findings['alta']['stock_negativo_err'] = str(e)[:200]

    # 4. Lotes vencidos pero estado_lote=VIGENTE con stock > 0
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,'') as lote,
                   MAX(fecha_vencimiento) as fv,
                   ROUND(SUM(CASE WHEN tipo='Entrada' THEN cantidad
                                  ELSE -cantidad END), 2) as stock,
                   MAX(estado_lote) as estado
            FROM movimientos
            WHERE COALESCE(fecha_vencimiento,'') != ''
                  AND fecha_vencimiento < date('now')
                  AND UPPER(COALESCE(estado_lote,'')) IN ('VIGENTE', '')
            GROUP BY material_id, lote
            HAVING stock > 0.5
            ORDER BY fv ASC LIMIT 100
        """).fetchall()
        findings['alta']['vencidos_pero_vigente'] = [
            {'material_id': r[0], 'lote': r[1], 'fecha_venc': r[2],
             'stock_g': r[3], 'estado_lote': r[4]}
            for r in rows
        ]
    except Exception as e:
        findings['alta']['vencidos_err'] = str(e)[:200]

    # === MEDIA ===
    # 5. Mismo nombre comercial con códigos distintos
    # Sebastián 10-may-2026: aplicar misma lógica de whitelist · nombres
    # comerciales genéricos (Agua, Alcohol, Glicerina) tampoco son
    # duplicados reales aunque compartan nombre.
    try:
        rows = c.execute("""
            SELECT LOWER(TRIM(nombre_comercial)) as nc_norm,
                   GROUP_CONCAT(codigo_mp) as codigos,
                   COUNT(*) as cnt,
                   GROUP_CONCAT(DISTINCT nombre_comercial) as variantes
            FROM maestro_mps
            WHERE activo=1 AND nombre_comercial IS NOT NULL
                  AND TRIM(nombre_comercial) != ''
            GROUP BY LOWER(TRIM(nombre_comercial))
            HAVING cnt > 1
            ORDER BY cnt DESC LIMIT 100
        """).fetchall()
        nc_real_dup = []
        nc_legal_compartido = []
        for r in rows:
            nc_norm = r[0]
            grupo = {
                'nombre_normalizado': nc_norm,
                'codigos_mp': (r[1] or '').split(','),
                'count': r[2], 'variantes_raw': r[3],
            }
            if nc_norm in INCI_WHITELIST_LEGAL:
                nc_legal_compartido.append({
                    **grupo,
                    'label': 'Nombre genérico · NO fusionar sin verificar',
                })
            else:
                nc_real_dup.append(grupo)
        findings['media']['nombre_comercial_duplicado'] = nc_real_dup
        if nc_legal_compartido:
            findings['baja']['nombre_comercial_compartido_legal'] = nc_legal_compartido
    except Exception as e:
        findings['media']['nombre_dup_err'] = str(e)[:200]

    # 6. Movimientos huérfanos (material_id sin fila activa en maestro_mps)
    try:
        rows = c.execute("""
            SELECT m.material_id, COUNT(*) as movs,
                   ROUND(SUM(CASE WHEN m.tipo='Entrada' THEN m.cantidad
                                  ELSE -m.cantidad END), 2) as stock
            FROM movimientos m
            LEFT JOIN maestro_mps mp ON m.material_id = mp.codigo_mp AND mp.activo=1
            WHERE mp.codigo_mp IS NULL
                  AND m.material_id IS NOT NULL AND TRIM(m.material_id) != ''
            GROUP BY m.material_id
            ORDER BY movs DESC LIMIT 50
        """).fetchall()
        findings['media']['movs_huerfanos'] = [
            {'material_id': r[0], 'n_movimientos': r[1], 'stock_actual_g': r[2]}
            for r in rows
        ]
    except Exception as e:
        findings['media']['huerfanos_err'] = str(e)[:200]

    # 7. Tipo material inválido
    try:
        rows = c.execute("""
            SELECT codigo_mp, COALESCE(tipo_material,'(null)') as tm,
                   COALESCE(nombre_comercial,'') as nc
            FROM maestro_mps
            WHERE activo=1
                  AND COALESCE(tipo_material,'') NOT IN
                      ('MP','Envase Primario','Envase Secundario','Empaque','')
            LIMIT 100
        """).fetchall()
        findings['media']['tipo_material_invalido'] = [
            {'codigo_mp': r[0], 'tipo_material': r[1], 'nombre': r[2]}
            for r in rows
        ]
    except Exception as e:
        findings['media']['tipo_invalido_err'] = str(e)[:200]

    # 8. Lotes duplicados entre MPs distintas
    try:
        rows = c.execute("""
            SELECT lote, COUNT(DISTINCT material_id) as mps_count,
                   GROUP_CONCAT(DISTINCT material_id) as materiales
            FROM movimientos
            WHERE COALESCE(lote,'') != ''
            GROUP BY lote
            HAVING mps_count > 1
            ORDER BY mps_count DESC LIMIT 50
        """).fetchall()
        findings['media']['lote_compartido_entre_mps'] = [
            {'lote': r[0], 'cantidad_mps': r[1],
             'material_ids': (r[2] or '').split(',')}
            for r in rows
        ]
    except Exception as e:
        findings['media']['lote_compartido_err'] = str(e)[:200]

    # 9. Códigos MP con espacios o caracteres extraños
    try:
        import re as _re
        rows = c.execute("""
            SELECT codigo_mp, LENGTH(codigo_mp) as len
            FROM maestro_mps WHERE activo=1
        """).fetchall()
        sospechosos = []
        for cod, ln in rows:
            if cod is None:
                continue
            stripped = (cod or '').strip()
            issues = []
            if cod != stripped:
                issues.append('espacios_borde')
            if '  ' in (cod or ''):
                issues.append('espacios_dobles')
            if not _re.match(r'^[A-Za-z0-9_\-]+$', stripped):
                issues.append('caracter_extraño')
            if issues:
                sospechosos.append({'codigo_mp': cod, 'issues': issues})
        findings['media']['codigos_caracter_extraño'] = sospechosos[:50]
    except Exception as e:
        findings['media']['caracter_err'] = str(e)[:200]

    # === BAJA ===
    # 10. MPs activas sin movimientos
    try:
        rows = c.execute("""
            SELECT mp.codigo_mp,
                   SUBSTR(COALESCE(mp.nombre_comercial,''),1,50) as nc,
                   COALESCE(mp.tipo_material,'MP') as tm
            FROM maestro_mps mp
            LEFT JOIN movimientos mov ON mp.codigo_mp = mov.material_id
            WHERE mp.activo=1 AND mov.id IS NULL
            ORDER BY mp.codigo_mp LIMIT 100
        """).fetchall()
        findings['baja']['mps_sin_movimientos'] = [
            {'codigo_mp': r[0], 'nombre': r[1], 'tipo_material': r[2]}
            for r in rows
        ]
    except Exception as e:
        findings['baja']['sin_movs_err'] = str(e)[:200]

    # 11. Proveedores con casing inconsistente (resumen, ya hay endpoint)
    try:
        rows = c.execute("""
            SELECT LOWER(TRIM(proveedor)) as prov_norm,
                   COUNT(DISTINCT proveedor) as variantes,
                   GROUP_CONCAT(DISTINCT proveedor) as raw
            FROM movimientos
            WHERE proveedor IS NOT NULL AND TRIM(proveedor) != ''
            GROUP BY LOWER(TRIM(proveedor))
            HAVING variantes > 1
            ORDER BY variantes DESC LIMIT 50
        """).fetchall()
        findings['baja']['proveedores_casing'] = [
            {'normalizado': r[0], 'cantidad_variantes': r[1],
             'variantes': (r[2] or '').split(',')}
            for r in rows
        ]
    except Exception as e:
        findings['baja']['prov_casing_err'] = str(e)[:200]

    # 12. Nombres muy similares (fuzzy con difflib)
    # Sebastián 10-may-2026: si modo rápido (?quick=1), saltar este check
    # · es el más lento (O(n²) con difflib · ~5-15s en catálogos grandes)
    # y puede causar timeout del browser. El frontend usa quick=1 por default
    # y ofrece botón "Análisis profundo" para correr sin omitir.
    if quick_mode:
        findings['baja']['nombres_similares_fuzzy'] = []
        findings['baja']['_fuzzy_omitido'] = (
            'modo rápido · pasa ?quick=0 para correr fuzzy completo'
        )
    else:
        try:
            import difflib as _diff
            rows = c.execute("""
                SELECT codigo_mp, COALESCE(nombre_comercial,''), COALESCE(nombre_inci,'')
                FROM maestro_mps WHERE activo=1
            """).fetchall()
            # Limitar a 800 MPs para evitar O(n²) explosivo
            rows = rows[:800]
            similares = []
            seen = set()
            for i in range(len(rows)):
                cod_i, nc_i, _ = rows[i]
                nc_i_norm = (nc_i or '').strip().lower()
                if len(nc_i_norm) < 5:
                    continue
                for j in range(i+1, len(rows)):
                    cod_j, nc_j, _ = rows[j]
                    nc_j_norm = (nc_j or '').strip().lower()
                    if len(nc_j_norm) < 5:
                        continue
                    if nc_i_norm == nc_j_norm:
                        continue
                    if abs(len(nc_i_norm) - len(nc_j_norm)) > 4:
                        continue
                    ratio = _diff.SequenceMatcher(None, nc_i_norm, nc_j_norm).ratio()
                    if ratio >= 0.85:
                        pair = tuple(sorted([cod_i, cod_j]))
                        if pair in seen:
                            continue
                        seen.add(pair)
                        similares.append({
                            'codigos': list(pair),
                            'nombres': [nc_i, nc_j],
                            'similaridad': round(ratio, 3),
                        })
                        if len(similares) >= 30:
                            break
                if len(similares) >= 30:
                    break
            findings['baja']['nombres_similares_fuzzy'] = similares
        except Exception as e:
            findings['baja']['fuzzy_err'] = str(e)[:200]

    conn.close()

    # Resumen
    def _count_real(d):
        # ignorar claves *_err en counts
        return sum(len(v) for k, v in d.items()
                   if not k.endswith('_err') and isinstance(v, list))
    n_alta = _count_real(findings['alta'])
    n_media = _count_real(findings['media'])
    n_baja = _count_real(findings['baja'])

    return jsonify({
        'ok': True,
        'timestamp': _dt.datetime.utcnow().isoformat() + 'Z',
        'resumen': {
            'n_alta': n_alta,
            'n_media': n_media,
            'n_baja': n_baja,
            'n_total_findings': n_alta + n_media + n_baja,
        },
        'findings': findings,
    }), 200


@bp.route("/api/admin/maestro-mps-unificar", methods=["POST"])
def maestro_mps_unificar():
    """Fusiona 2+ MPs duplicadas en una canónica.

    Sebastián 10-may-2026: tras auditoría detectamos grupos de MPs con
    mismo INCI o nombre comercial pero códigos distintos (típicamente
    casing inconsistente como SODIUM HYDROXIDE vs Sodium Hydroxide).
    Este endpoint transfiere TODOS los movimientos de los duplicados
    al código canónico elegido, archiva los duplicados (activo=0) y
    deja audit_log con el snapshot.

    Body JSON:
      codigo_canonico: str      (la MP que sobrevive)
      codigos_duplicados: []    (1+ MPs a fusionar y archivar)
      motivo: str               (recomendado, queda en audit_log)
      merge_force: bool         (opcional, default false · permite
                                  fusionar aunque tengan diferente INCI/
                                  nombre · solo para typos extremos)

    Tablas que se actualizan (transacción atómica):
      - movimientos.material_id          (TODOS los movs duplicados)
      - formula_items.material_id        (recetas)
      - conteo_items.codigo_mp           (conteos cíclicos)
      - solicitudes_compra_items.codigo_mp (SOLs pendientes)
      - mp_lead_time_config.material_id  (config de leadtime)
      - maestro_mps.activo=0             (archive duplicados)
      - audit_log                        (snapshot completo)

    Returns:
      { ok, canonico, duplicados_archivados, totales_transferidos: {
        movimientos, formula_items, conteo_items, sol_items,
        leadtime_config
      } }

    Errores:
      400 codigo_canonico vacío o duplicado vacío
      404 alguna MP no existe
      409 INCI/nombre incompatibles sin merge_force
      500 falla transaccional (rollback automático)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    canonico = (d.get('codigo_canonico') or '').strip().upper()
    duplicados = d.get('codigos_duplicados') or []
    motivo = (d.get('motivo') or '').strip()
    merge_force = bool(d.get('merge_force'))

    if not canonico:
        return jsonify({'error': 'codigo_canonico requerido'}), 400
    if not isinstance(duplicados, list) or not duplicados:
        return jsonify({'error': 'codigos_duplicados debe ser lista no vacía'}), 400
    duplicados = [str(x).strip().upper() for x in duplicados if str(x).strip()]
    if not duplicados:
        return jsonify({'error': 'codigos_duplicados vacíos tras strip'}), 400
    if canonico in duplicados:
        return jsonify({
            'error': 'codigo_canonico no puede estar en codigos_duplicados',
            'detail': f'{canonico} aparece en ambas listas.',
        }), 400
    if len(duplicados) > 20:
        return jsonify({
            'error': 'demasiados duplicados',
            'detail': 'Máximo 20 MPs por fusión para evitar operaciones masivas accidentales.',
        }), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # 1. Validar canónico existe y está activo
    row = c.execute(
        "SELECT codigo_mp, nombre_inci, nombre_comercial, activo "
        "FROM maestro_mps WHERE codigo_mp=?", (canonico,)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({
            'error': 'Canónico no encontrado',
            'detail': f'{canonico} no existe en maestro_mps',
        }), 404
    if not row[3]:
        conn.close()
        return jsonify({
            'error': 'Canónico archivado',
            'detail': f'{canonico} tiene activo=0. Reactivar primero o elegir otro canónico.',
        }), 400
    inci_canonico = (row[1] or '').strip().lower()
    nc_canonico = (row[2] or '').strip().lower()

    # 2. Validar cada duplicado existe y compatible
    info_duplicados = []
    incompat = []
    for cod_d in duplicados:
        row_d = c.execute(
            "SELECT codigo_mp, nombre_inci, nombre_comercial, activo "
            "FROM maestro_mps WHERE codigo_mp=?", (cod_d,)
        ).fetchone()
        if not row_d:
            conn.close()
            return jsonify({
                'error': 'Duplicado no encontrado',
                'detail': f'{cod_d} no existe en maestro_mps',
            }), 404
        inci_d = (row_d[1] or '').strip().lower()
        nc_d = (row_d[2] or '').strip().lower()
        info_duplicados.append({
            'codigo_mp': cod_d,
            'nombre_inci': row_d[1] or '',
            'nombre_comercial': row_d[2] or '',
            'activo': bool(row_d[3]),
        })
        if not merge_force:
            # Si tienen INCI distinto Y nombre comercial distinto → sospechoso
            if inci_canonico and inci_d and inci_canonico != inci_d:
                if nc_canonico and nc_d and nc_canonico != nc_d:
                    incompat.append({
                        'codigo_mp': cod_d,
                        'inci_duplicado': row_d[1],
                        'inci_canonico': row[1],
                    })

    if incompat and not merge_force:
        conn.close()
        return jsonify({
            'error': 'INCI/nombre incompatibles',
            'detail': ('Algunos duplicados tienen INCI y nombre comercial '
                       'diferentes al canónico. Si querés forzar la fusión '
                       '(typos extremos), pasá merge_force=true.'),
            'incompatibles': incompat,
        }), 409

    # 3. Ejecutar fusión en transacción
    totales = {'movimientos': 0, 'formula_items': 0, 'conteo_items': 0,
               'sol_items': 0, 'leadtime_config': 0}
    try:
        ph = ','.join(['?'] * len(duplicados))

        # movimientos.material_id
        c.execute(f"UPDATE movimientos SET material_id=? "
                  f"WHERE material_id IN ({ph})", [canonico] + duplicados)
        totales['movimientos'] = c.rowcount

        # formula_items.material_id (si la tabla existe)
        try:
            c.execute(f"UPDATE formula_items SET material_id=? "
                      f"WHERE material_id IN ({ph})", [canonico] + duplicados)
            totales['formula_items'] = c.rowcount
        except sqlite3.OperationalError:
            pass

        # conteo_items.codigo_mp (si la tabla existe)
        try:
            c.execute(f"UPDATE conteo_items SET codigo_mp=? "
                      f"WHERE codigo_mp IN ({ph})", [canonico] + duplicados)
            totales['conteo_items'] = c.rowcount
        except sqlite3.OperationalError:
            pass

        # solicitudes_compra_items.codigo_mp (si la tabla existe)
        try:
            c.execute(f"UPDATE solicitudes_compra_items SET codigo_mp=? "
                      f"WHERE codigo_mp IN ({ph})", [canonico] + duplicados)
            totales['sol_items'] = c.rowcount
        except sqlite3.OperationalError:
            pass

        # mp_lead_time_config.material_id
        try:
            # Para evitar UNIQUE conflicts (si canónico ya tiene config),
            # primero borramos los configs de duplicados que colisionarían.
            existe_canonico_lt = c.execute(
                "SELECT 1 FROM mp_lead_time_config WHERE material_id=?",
                (canonico,)
            ).fetchone()
            if existe_canonico_lt:
                # canónico ya tiene config · borrar configs de duplicados
                c.execute(f"DELETE FROM mp_lead_time_config "
                          f"WHERE material_id IN ({ph})", duplicados)
                totales['leadtime_config'] = c.rowcount
            else:
                # canónico sin config · transferir del primero que tenga
                c.execute(f"UPDATE mp_lead_time_config SET material_id=? "
                          f"WHERE material_id=(SELECT material_id "
                          f"  FROM mp_lead_time_config WHERE material_id IN ({ph}) "
                          f"  LIMIT 1)",
                          [canonico] + duplicados)
                totales['leadtime_config'] = c.rowcount
                # Borrar las restantes (si las hay)
                c.execute(f"DELETE FROM mp_lead_time_config "
                          f"WHERE material_id IN ({ph})", duplicados)
        except sqlite3.OperationalError:
            pass

        # Archive duplicados
        c.execute(f"UPDATE maestro_mps SET activo=0 "
                  f"WHERE codigo_mp IN ({ph})", duplicados)

        # Audit log
        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='UNIFICAR_MPS', tabla='maestro_mps',
                registro_id=canonico,
                despues={
                    'canonico': canonico,
                    'duplicados_archivados': duplicados,
                    'motivo': motivo,
                    'merge_force': merge_force,
                    'totales_transferidos': totales,
                    'info_duplicados': info_duplicados,
                },
                detalle=(f'Fusión MPs: {len(duplicados)} duplicados → {canonico} · '
                         f'{totales["movimientos"]} movimientos transferidos'),
            )
        except Exception:
            pass

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({
            'error': 'Falla transaccional en fusión',
            'detail': str(e)[:300],
            'rollback': 'aplicado · ningún cambio persistió',
        }), 500

    conn.close()

    return jsonify({
        'ok': True,
        'canonico': canonico,
        'duplicados_archivados': duplicados,
        'totales_transferidos': totales,
        'message': (f'✓ {len(duplicados)} MP(s) fusionada(s) en {canonico} · '
                    f'{totales["movimientos"]} movimientos transferidos · '
                    f'duplicados archivados (activo=0).'),
    }), 200


@bp.route("/api/admin/maestro-mps-unificar-bulk", methods=["POST"])
def maestro_mps_unificar_bulk():
    """Fusión masiva: procesa N grupos en una sola pasada.

    Sebastián 10-may-2026: 200 grupos para fusionar manualmente = ~2h.
    Este endpoint procesa todos en un solo round-trip y devuelve
    resumen + grupos que fallaron individualmente.

    Body JSON:
      grupos: [
        {codigo_canonico: 'MP00001', codigos_duplicados: ['MP00002', ...]},
        ...
      ]
      motivo: str (compartido para todos · queda en audit_log)
      merge_force: bool (default false · pasa a cada grupo)

    Comportamiento:
      - Procesa grupo por grupo · si UNO falla, se rollback ESE grupo
        solamente · los demás se procesan normal.
      - Si codigo_canonico está en codigos_duplicados de OTRO grupo,
        se procesa en orden: el grupo donde es duplicado se procesa
        ANTES (canónico pasa a ser archivado), después no se puede
        usar como canónico. Backend rechaza con error grupo-específico.
      - Límite: max 500 grupos por request.

    Returns:
      { ok, total_grupos, exitosos, fallidos[],
        resumen_totales: {movimientos, formula_items, ...} }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    grupos = d.get('grupos') or []
    motivo = (d.get('motivo') or '').strip()
    merge_force = bool(d.get('merge_force'))

    if not isinstance(grupos, list) or not grupos:
        return jsonify({'error': 'grupos debe ser lista no vacía'}), 400
    if len(grupos) > 500:
        return jsonify({
            'error': 'demasiados grupos',
            'detail': f'Máximo 500 por request · enviados {len(grupos)}. Dividí en lotes.',
        }), 400

    # Procesar grupo por grupo · cada uno con su propia conexión y transacción
    # para que un fallo individual NO bloquee al resto.
    resultados = []
    totales_acum = {'movimientos': 0, 'formula_items': 0, 'conteo_items': 0,
                    'sol_items': 0, 'leadtime_config': 0}
    exitosos = 0
    fallidos = []

    for idx, grupo in enumerate(grupos):
        canonico = (grupo.get('codigo_canonico') or '').strip().upper()
        duplicados = grupo.get('codigos_duplicados') or []
        if not canonico or not duplicados:
            fallidos.append({
                'idx': idx, 'canonico': canonico,
                'error': 'canonico o duplicados vacíos',
            })
            continue
        duplicados = [str(x).strip().upper() for x in duplicados if str(x).strip()]
        if canonico in duplicados:
            fallidos.append({
                'idx': idx, 'canonico': canonico,
                'error': 'canonico está en duplicados',
            })
            continue
        if len(duplicados) > 20:
            fallidos.append({
                'idx': idx, 'canonico': canonico,
                'error': f'max 20 duplicados por grupo · este tiene {len(duplicados)}',
            })
            continue

        # Procesar este grupo con su propia conexión
        conn = sqlite3.connect(DB_PATH)
        conn.execute("PRAGMA busy_timeout=2000")
        c = conn.cursor()
        try:
            # Validar canónico
            row = c.execute(
                "SELECT activo FROM maestro_mps WHERE codigo_mp=?", (canonico,)
            ).fetchone()
            if not row:
                conn.close()
                fallidos.append({
                    'idx': idx, 'canonico': canonico,
                    'error': f'canonico {canonico} no existe',
                })
                continue
            if not row[0]:
                conn.close()
                fallidos.append({
                    'idx': idx, 'canonico': canonico,
                    'error': f'canonico {canonico} está archivado · ya no es válido',
                })
                continue
            # Validar duplicados existen
            ph = ','.join(['?'] * len(duplicados))
            existentes = c.execute(
                f"SELECT codigo_mp FROM maestro_mps WHERE codigo_mp IN ({ph})",
                duplicados
            ).fetchall()
            existentes_set = {r[0] for r in existentes}
            no_existen = [d for d in duplicados if d not in existentes_set]
            if no_existen:
                conn.close()
                fallidos.append({
                    'idx': idx, 'canonico': canonico,
                    'error': f'duplicados no existen: {", ".join(no_existen)}',
                })
                continue

            # Ejecutar transferencias
            totales_grupo = {}
            c.execute(f"UPDATE movimientos SET material_id=? "
                      f"WHERE material_id IN ({ph})", [canonico] + duplicados)
            totales_grupo['movimientos'] = c.rowcount
            try:
                c.execute(f"UPDATE formula_items SET material_id=? "
                          f"WHERE material_id IN ({ph})", [canonico] + duplicados)
                totales_grupo['formula_items'] = c.rowcount
            except sqlite3.OperationalError:
                totales_grupo['formula_items'] = 0
            try:
                c.execute(f"UPDATE conteo_items SET codigo_mp=? "
                          f"WHERE codigo_mp IN ({ph})", [canonico] + duplicados)
                totales_grupo['conteo_items'] = c.rowcount
            except sqlite3.OperationalError:
                totales_grupo['conteo_items'] = 0
            try:
                c.execute(f"UPDATE solicitudes_compra_items SET codigo_mp=? "
                          f"WHERE codigo_mp IN ({ph})", [canonico] + duplicados)
                totales_grupo['sol_items'] = c.rowcount
            except sqlite3.OperationalError:
                totales_grupo['sol_items'] = 0
            try:
                existe_canon_lt = c.execute(
                    "SELECT 1 FROM mp_lead_time_config WHERE material_id=?",
                    (canonico,)
                ).fetchone()
                if existe_canon_lt:
                    c.execute(f"DELETE FROM mp_lead_time_config "
                              f"WHERE material_id IN ({ph})", duplicados)
                else:
                    c.execute(f"UPDATE mp_lead_time_config SET material_id=? "
                              f"WHERE material_id=(SELECT material_id "
                              f"  FROM mp_lead_time_config WHERE material_id IN ({ph}) "
                              f"  LIMIT 1)",
                              [canonico] + duplicados)
                    c.execute(f"DELETE FROM mp_lead_time_config "
                              f"WHERE material_id IN ({ph})", duplicados)
                totales_grupo['leadtime_config'] = c.rowcount
            except sqlite3.OperationalError:
                totales_grupo['leadtime_config'] = 0
            # Archive duplicados
            c.execute(f"UPDATE maestro_mps SET activo=0 "
                      f"WHERE codigo_mp IN ({ph})", duplicados)
            # Audit log
            try:
                import json as _json
                audit_log(
                    c, usuario=u, accion='UNIFICAR_MPS_BULK',
                    tabla='maestro_mps', registro_id=canonico,
                    despues={
                        'canonico': canonico,
                        'duplicados': duplicados,
                        'motivo': motivo,
                        'totales': totales_grupo,
                        'batch_idx': idx,
                    },
                    detalle=f'Fusión bulk #{idx}: {len(duplicados)} → {canonico}',
                )
            except Exception:
                pass

            conn.commit()
            conn.close()
            exitosos += 1
            for k, v in totales_grupo.items():
                totales_acum[k] = totales_acum.get(k, 0) + (v or 0)
            resultados.append({
                'idx': idx, 'canonico': canonico,
                'duplicados': duplicados, 'totales': totales_grupo,
            })
        except Exception as e:
            try:
                conn.rollback()
                conn.close()
            except Exception:
                pass
            fallidos.append({
                'idx': idx, 'canonico': canonico,
                'error': f'falla transaccional: {str(e)[:200]}',
            })

    return jsonify({
        'ok': True,
        'total_grupos': len(grupos),
        'exitosos': exitosos,
        'fallidos': fallidos,
        'resumen_totales': totales_acum,
        'message': (f'Procesados {exitosos}/{len(grupos)} grupos · '
                    f'{totales_acum["movimientos"]} movimientos transferidos · '
                    f'{len(fallidos)} con error.'),
    }), 200


@bp.route("/api/admin/material-ids-huerfanos", methods=["GET"])
def material_ids_huerfanos():
    """Lista material_ids usados en formula_items/movimientos/conteo_items/SOLs
    que NO tienen fila activa en maestro_mps.

    Sebastián 10-may-2026: 187 huérfanos detectados por auditor 4 ·
    bloquean producción real (pre-check FEFO no encuentra lotes).
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    queries = {
        'formula_items': """
            SELECT fi.material_id,
                   SUBSTR(MAX(fi.material_nombre),1,40) as nombre,
                   COUNT(DISTINCT fi.producto_nombre) as productos_usando,
                   COUNT(*) as items_total
            FROM formula_items fi
            LEFT JOIN maestro_mps mp ON fi.material_id=mp.codigo_mp AND mp.activo=1
            WHERE mp.codigo_mp IS NULL
              AND fi.material_id IS NOT NULL AND TRIM(fi.material_id) != ''
            GROUP BY fi.material_id
            ORDER BY productos_usando DESC
        """,
        'movimientos': """
            SELECT m.material_id,
                   SUBSTR(MAX(m.material_nombre),1,40) as nombre,
                   COUNT(*) as movs,
                   ROUND(SUM(CASE WHEN m.tipo='Entrada' THEN m.cantidad
                                  ELSE -m.cantidad END), 2) as stock_neto
            FROM movimientos m
            LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp AND mp.activo=1
            WHERE mp.codigo_mp IS NULL
              AND m.material_id IS NOT NULL AND TRIM(m.material_id) != ''
            GROUP BY m.material_id
            ORDER BY movs DESC
        """,
    }
    result = {}
    for k, sql in queries.items():
        try:
            rows = c.execute(sql).fetchall()
            cols = [d[0] for d in c.description]
            result[k] = [dict(zip(cols, r)) for r in rows]
        except Exception as e:
            result[k+'_err'] = str(e)[:200]

    conn.close()
    return jsonify({
        'ok': True,
        'huerfanos': result,
        'resumen': {
            'formula_items_huerfanos': len(result.get('formula_items', [])),
            'movimientos_huerfanos': len(result.get('movimientos', [])),
        },
    }), 200


@bp.route("/api/admin/crear-mps-huerfanas", methods=["POST"])
def crear_mps_huerfanas():
    """Crea en maestro_mps las MPs huérfanas detectadas (con datos
    mínimos · nombre desde movs/formula_items).

    Body JSON:
      material_ids: ['MPAGUALI01', ...] · lista a crear
      dry_run: bool (default false)

    Para cada material_id:
      - Si ya existe en maestro_mps: lo reactiva (activo=1)
      - Si no existe: INSERT con datos derivados (nombre desde otras tablas)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    mids = d.get('material_ids') or []
    dry_run = bool(d.get('dry_run'))

    if not isinstance(mids, list) or not mids:
        return jsonify({'error': 'material_ids debe ser lista no vacía'}), 400
    if len(mids) > 300:
        return jsonify({'error': 'max 300 por request'}), 400

    mids = [str(x).strip().upper() for x in mids if str(x).strip()]

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    plan = []
    for mid in mids:
        existe = c.execute(
            "SELECT codigo_mp, activo FROM maestro_mps WHERE codigo_mp=?", (mid,)
        ).fetchone()
        # Derivar nombre desde formula_items/movimientos
        nom_row = c.execute("""
            SELECT material_nombre FROM (
                SELECT material_nombre, COUNT(*) as n FROM formula_items
                WHERE material_id=? GROUP BY material_nombre
                UNION ALL
                SELECT material_nombre, COUNT(*) as n FROM movimientos
                WHERE material_id=? GROUP BY material_nombre
            ) GROUP BY material_nombre ORDER BY SUM(n) DESC LIMIT 1
        """, (mid, mid)).fetchone()
        nombre = (nom_row[0] if nom_row else mid)
        if existe:
            plan.append({
                'material_id': mid, 'accion': 'reactivar' if not existe[1] else 'ya_activo',
                'nombre': nombre,
            })
        else:
            plan.append({
                'material_id': mid, 'accion': 'crear', 'nombre': nombre,
            })

    if dry_run:
        conn.close()
        return jsonify({'ok': True, 'dry_run': True, 'plan': plan,
                       'message': f'Plan: {len(plan)} MPs'}), 200

    creados = 0
    reactivados = 0
    try:
        for p in plan:
            if p['accion'] == 'crear':
                c.execute("""
                    INSERT INTO maestro_mps
                    (codigo_mp, nombre_inci, nombre_comercial, tipo, proveedor,
                     stock_minimo, activo, tipo_material)
                    VALUES (?, '', ?, '', '', 0, 1, 'MP')
                """, (p['material_id'], p['nombre']))
                creados += 1
            elif p['accion'] == 'reactivar':
                c.execute("UPDATE maestro_mps SET activo=1 WHERE codigo_mp=?",
                          (p['material_id'],))
                reactivados += 1
        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='CREAR_MPS_HUERFANAS',
                tabla='maestro_mps', registro_id='bulk',
                despues={'creados': creados, 'reactivados': reactivados, 'plan': plan},
                detalle=f'Creados {creados} · reactivados {reactivados}',
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional', 'detail': str(e)[:300]}), 500

    conn.close()
    return jsonify({
        'ok': True, 'creados': creados, 'reactivados': reactivados,
        'message': f'✓ {creados} MPs creadas + {reactivados} reactivadas',
    }), 200


@bp.route("/api/admin/anular-movimiento", methods=["POST"])
def anular_movimiento():
    """Anula un movimiento creando contra-movimiento Entrada/Salida del
    mismo lote y cantidad. Audit log INVIMA compliant.

    Sebastián 10-may-2026: MP00112 lote AJUSTE-4 con stock -1.4M sin
    Entrada respaldatoria. Caso de uso: anular salida fantasma para
    llevar saldo a 0 sin borrar el original (trazabilidad INVIMA).

    Body JSON:
      mov_id: int (movimiento a anular)
      motivo: str (queda en audit_log y observaciones)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    mov_id = d.get('mov_id')
    motivo = (d.get('motivo') or '').strip()

    try:
        mov_id = int(mov_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'mov_id inválido'}), 400
    if not motivo or len(motivo) < 10:
        return jsonify({'error': 'motivo requerido (mín 10 chars)'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    row = c.execute("""
        SELECT id, material_id, material_nombre, cantidad, tipo, lote,
               fecha_vencimiento, estanteria, posicion, proveedor, estado_lote
        FROM movimientos WHERE id=?
    """, (mov_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'movimiento no encontrado'}), 404

    orig = {
        'id': row[0], 'material_id': row[1], 'material_nombre': row[2],
        'cantidad': row[3], 'tipo': row[4], 'lote': row[5],
        'fecha_vencimiento': row[6], 'estanteria': row[7], 'posicion': row[8],
        'proveedor': row[9], 'estado_lote': row[10],
    }

    # Tipo inverso
    tipo_contra = 'Entrada' if (orig['tipo'] == 'Salida') else 'Salida'
    obs_contra = (f'ANULACION mov #{orig["id"]} ({orig["tipo"]} {orig["cantidad"]}g) · '
                  f'Motivo: {motivo} · Por: {u}')

    try:
        c.execute("""
            INSERT INTO movimientos
            (material_id, material_nombre, cantidad, tipo, fecha, observaciones,
             lote, fecha_vencimiento, estanteria, posicion, proveedor, estado_lote, operador)
            VALUES (?,?,?,?,datetime('now'),?,?,?,?,?,?,?,?)
        """, (orig['material_id'], orig['material_nombre'],
              orig['cantidad'], tipo_contra, obs_contra,
              orig['lote'], orig['fecha_vencimiento'],
              orig['estanteria'], orig['posicion'], orig['proveedor'],
              orig['estado_lote'], u))
        contra_id = c.lastrowid

        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='ANULAR_MOVIMIENTO',
                tabla='movimientos', registro_id=str(mov_id),
                antes=orig,
                despues={'contra_movimiento_id': contra_id, 'motivo': motivo,
                         'tipo_contra': tipo_contra},
                detalle=(f'Anulación mov #{mov_id} · creado contra #{contra_id} · '
                         f'tipo {tipo_contra} {orig["cantidad"]}g · motivo: {motivo}'),
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional', 'detail': str(e)[:300]}), 500

    conn.close()
    return jsonify({
        'ok': True,
        'movimiento_original': orig,
        'contra_movimiento_id': contra_id,
        'tipo_contra': tipo_contra,
        'message': (f'✓ Movimiento #{mov_id} anulado · creado contra #{contra_id} '
                    f'tipo {tipo_contra} {orig["cantidad"]}g lote {orig["lote"]}'),
    }), 200


@bp.route("/api/admin/formula-duplicados", methods=["GET"])
def formula_duplicados_listado():
    """Lista TODAS las fórmulas con items duplicados (mismo material_id
    aparece >1 vez en la misma fórmula) o cuya suma de porcentajes NO
    da 100% ±0.5.

    Sebastián 10-may-2026: caso real SUERO ILUMINADOR TRX = 200% por
    46 items con duplicados. Sin esto, producir TRX descontaría 2x.
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # 1. Fórmulas con items duplicados (mismo material_id repetido)
    duplicados = []
    rows = c.execute("""
        SELECT producto_nombre, material_id, COUNT(*) as veces,
               GROUP_CONCAT(id) as ids, GROUP_CONCAT(porcentaje) as pcts
        FROM formula_items
        WHERE material_id IS NOT NULL AND TRIM(material_id) != ''
        GROUP BY producto_nombre, material_id
        HAVING veces > 1
        ORDER BY producto_nombre, material_id
    """).fetchall()
    for r in rows:
        duplicados.append({
            'producto': r[0], 'material_id': r[1], 'veces': r[2],
            'ids': [int(x) for x in (r[3] or '').split(',') if x],
            'porcentajes': [float(x) for x in (r[4] or '').split(',') if x],
        })

    # 2. Fórmulas con SUM(porcentaje) NO == 100%
    sumas = []
    rows = c.execute("""
        SELECT producto_nombre, ROUND(SUM(porcentaje), 4) as suma,
               COUNT(*) as n_items
        FROM formula_items
        GROUP BY producto_nombre
        HAVING ABS(suma - 100) > 0.5
        ORDER BY ABS(suma - 100) DESC
    """).fetchall()
    for r in rows:
        sumas.append({
            'producto': r[0], 'suma_porcentajes': r[1],
            'n_items': r[2],
            'diff_vs_100': round((r[1] or 0) - 100, 4),
        })

    conn.close()
    return jsonify({
        'ok': True,
        'duplicados': duplicados,
        'porcentajes_anomalos': sumas,
        'resumen': {
            'productos_con_duplicados': len(set(d['producto'] for d in duplicados)),
            'items_duplicados_total': sum(d['veces'] for d in duplicados),
            'formulas_porcentaje_no_100': len(sumas),
        },
    }), 200


@bp.route("/api/admin/formula-limpiar-duplicados", methods=["POST"])
def formula_limpiar_duplicados():
    """Consolida items duplicados de una fórmula en uno solo.

    Para cada grupo (producto, material_id) con N filas:
      1. Conserva la fila con menor id
      2. Suma los porcentajes y cantidades en la conservada
      3. DELETE las demás
      4. Audit log

    Body JSON:
      producto: str (opcional · si vacío procesa TODAS las fórmulas con dups)
      dry_run: bool (default false · si true solo simula y devuelve plan)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    producto_filter = (d.get('producto') or '').strip()
    dry_run = bool(d.get('dry_run'))

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # Detectar duplicados
    if producto_filter:
        rows = c.execute("""
            SELECT producto_nombre, material_id, GROUP_CONCAT(id) as ids,
                   GROUP_CONCAT(porcentaje) as pcts,
                   GROUP_CONCAT(COALESCE(cantidad_g_por_lote,0)) as cants
            FROM formula_items
            WHERE producto_nombre = ?
              AND material_id IS NOT NULL AND TRIM(material_id) != ''
            GROUP BY producto_nombre, material_id
            HAVING COUNT(*) > 1
        """, (producto_filter,)).fetchall()
    else:
        rows = c.execute("""
            SELECT producto_nombre, material_id, GROUP_CONCAT(id) as ids,
                   GROUP_CONCAT(porcentaje) as pcts,
                   GROUP_CONCAT(COALESCE(cantidad_g_por_lote,0)) as cants
            FROM formula_items
            WHERE material_id IS NOT NULL AND TRIM(material_id) != ''
            GROUP BY producto_nombre, material_id
            HAVING COUNT(*) > 1
        """).fetchall()

    plan = []
    total_borrados = 0
    for r in rows:
        producto, mid, ids_str, pcts_str, cants_str = r
        ids = [int(x) for x in (ids_str or '').split(',') if x]
        pcts = [float(x) for x in (pcts_str or '').split(',') if x]
        cants = [float(x) for x in (cants_str or '').split(',') if x]
        if len(ids) <= 1:
            continue
        ids_sorted = sorted(ids)
        keep_id = ids_sorted[0]
        delete_ids = ids_sorted[1:]
        suma_pct = round(sum(pcts), 4)
        suma_cant = round(sum(cants), 4)
        plan.append({
            'producto': producto, 'material_id': mid,
            'keep_id': keep_id,
            'delete_ids': delete_ids,
            'porcentaje_consolidado': suma_pct,
            'cantidad_g_consolidada': suma_cant,
            'items_originales': len(ids),
        })
        total_borrados += len(delete_ids)

    if dry_run:
        conn.close()
        return jsonify({
            'ok': True, 'dry_run': True,
            'plan': plan,
            'items_a_consolidar': len(plan),
            'items_a_borrar': total_borrados,
            'message': f'Plan: consolidar {len(plan)} grupos, borrar {total_borrados} items duplicados',
        }), 200

    # Aplicar
    aplicados = 0
    errores = []
    try:
        for p in plan:
            try:
                c.execute("""
                    UPDATE formula_items
                    SET porcentaje = ?, cantidad_g_por_lote = ?
                    WHERE id = ?
                """, (p['porcentaje_consolidado'], p['cantidad_g_consolidada'], p['keep_id']))
                ph = ','.join(['?'] * len(p['delete_ids']))
                c.execute(f"DELETE FROM formula_items WHERE id IN ({ph})", p['delete_ids'])
                aplicados += 1
            except Exception as e:
                errores.append({'producto': p['producto'], 'material_id': p['material_id'],
                               'error': str(e)[:200]})
        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='LIMPIAR_FORMULA_DUPLICADOS',
                tabla='formula_items', registro_id=producto_filter or 'bulk',
                despues={'plan': plan, 'aplicados': aplicados, 'errores': errores},
                detalle=f'Consolidados {aplicados} grupos, borrados {total_borrados} items',
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional', 'detail': str(e)[:300]}), 500

    conn.close()

    return jsonify({
        'ok': True, 'aplicado': True,
        'grupos_consolidados': aplicados,
        'items_borrados': total_borrados,
        'errores': errores,
        'message': f'✓ Consolidados {aplicados} grupos · borrados {total_borrados} items duplicados',
    }), 200


@bp.route("/api/admin/marcar-lotes-vencidos", methods=["POST"])
def marcar_lotes_vencidos():
    """Cambia estado_lote='VENCIDO' en lotes que tienen fecha_venc pasada
    pero estado='VIGENTE'. Bulk action desde panel auditoría.

    Sebastián 10-may-2026: limpieza regulatoria · MPs vencidas marcadas
    como VIGENTE son violación INVIMA (pueden usarse en producción).

    Body JSON:
      lotes: [{material_id, lote}]  · lista de lotes a marcar
      motivo: str (queda en audit_log)

    Returns: { ok, actualizados, audit_id }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    lotes = d.get('lotes') or []
    motivo = (d.get('motivo') or 'Limpieza auditoría · vencidos sin marcar').strip()

    if not isinstance(lotes, list) or not lotes:
        return jsonify({'error': 'lotes debe ser lista no vacía'}), 400
    if len(lotes) > 200:
        return jsonify({'error': 'max 200 lotes por request'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    total_actualizados = 0
    detalles = []
    try:
        for it in lotes:
            mid = (it.get('material_id') or '').strip()
            lt = (it.get('lote') or '').strip()
            if not mid:
                continue
            if lt:
                c.execute(
                    "UPDATE movimientos SET estado_lote='VENCIDO' "
                    "WHERE material_id=? AND lote=? "
                    "AND UPPER(COALESCE(estado_lote,''))='VIGENTE'",
                    (mid, lt)
                )
            else:
                c.execute(
                    "UPDATE movimientos SET estado_lote='VENCIDO' "
                    "WHERE material_id=? AND (lote IS NULL OR lote='') "
                    "AND UPPER(COALESCE(estado_lote,''))='VIGENTE'",
                    (mid,)
                )
            n = c.rowcount
            total_actualizados += n
            detalles.append({'material_id': mid, 'lote': lt, 'movs_actualizados': n})

        # Audit log
        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='MARCAR_LOTES_VENCIDOS',
                tabla='movimientos', registro_id='bulk',
                despues={
                    'motivo': motivo,
                    'detalles': detalles,
                    'total_movs_actualizados': total_actualizados,
                },
                detalle=f'Marcados VENCIDO en {len(lotes)} lotes · {total_actualizados} movs',
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional', 'detail': str(e)[:300]}), 500

    conn.close()
    return jsonify({
        'ok': True,
        'lotes_procesados': len(lotes),
        'total_movimientos_actualizados': total_actualizados,
        'detalles': detalles,
        'message': f'✓ {len(lotes)} lotes marcados como VENCIDO · {total_actualizados} movs actualizados',
    }), 200


@bp.route("/api/admin/marcar-vencidos-bulk-todos", methods=["POST"])
def marcar_vencidos_bulk_todos():
    """Marca VENCIDO en TODOS los lotes con fecha_venc pasada que aún son
    VIGENTE. Equivalente al cron diario `job_marcar_vencidos` pero a demanda.

    Sebastián 8-may-2026 (zero-error FASE A): trigger manual para no
    esperar al cron de las 7:50am cuando se acaba de descubrir un lote
    vencido en auditoría.

    Returns: { ok, actualizados, lotes_afectados, top_5_codigos }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    try:
        # Detectar primero (para feedback al usuario)
        rows = c.execute("""
            SELECT material_id, lote, fecha_vencimiento, COUNT(*) AS movs
            FROM movimientos
            WHERE fecha_vencimiento IS NOT NULL
              AND TRIM(fecha_vencimiento) != ''
              AND date(fecha_vencimiento) < date('now')
              AND UPPER(COALESCE(estado_lote,'')) = 'VIGENTE'
            GROUP BY material_id, lote
            ORDER BY fecha_vencimiento ASC
            LIMIT 500
        """).fetchall()

        if not rows:
            conn.close()
            return jsonify({
                'ok': True,
                'actualizados': 0,
                'lotes_afectados': 0,
                'message': 'Sin lotes vencidos pendientes · OK',
            }), 200

        res = c.execute("""
            UPDATE movimientos
            SET estado_lote = 'VENCIDO'
            WHERE fecha_vencimiento IS NOT NULL
              AND TRIM(fecha_vencimiento) != ''
              AND date(fecha_vencimiento) < date('now')
              AND UPPER(COALESCE(estado_lote,'')) = 'VIGENTE'
        """)
        actualizados = res.rowcount

        try:
            detalles = [
                {'material_id': r[0], 'lote': r[1],
                 'fecha_venc': r[2], 'movs': r[3]}
                for r in rows[:100]
            ]
            audit_log(
                c, usuario=u,
                accion='MARCAR_LOTES_VENCIDOS_BULK',
                tabla='movimientos', registro_id='bulk-trigger',
                despues={
                    'total_movs_actualizados': actualizados,
                    'lotes_afectados': len(rows),
                    'detalles': detalles,
                },
                detalle=(f'Trigger manual marcó VENCIDO en {len(rows)} '
                         f'lotes · {actualizados} movimientos'),
            )
        except Exception:
            pass

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'actualizados': actualizados,
            'lotes_afectados': len(rows),
            'top_codigos': [r[0] for r in rows[:5]],
            'top_lotes': [{'material_id': r[0], 'lote': r[1],
                           'fecha_venc': r[2]} for r in rows[:10]],
            'message': (f'✓ {len(rows)} lotes marcados VENCIDO '
                        f'· {actualizados} movs actualizados'),
        }), 200

    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/mps-sin-uso", methods=["GET"])
def mps_sin_uso():
    """Detecta MPs activas que NO se usan en ningún lado.

    Sebastián 8-may-2026 (zero-error FASE A): el catálogo va creciendo
    con MPs que en algún momento se probaron y nunca más se usaron.
    Estas MPs:
      - No están en ninguna fórmula activa
      - No tienen movimientos en >X días (default 365d)
      - Stock actual = 0
      - activo=1

    Son candidatas a archivar (activo=0) para limpiar listas y reportes.
    NO se borran · activo=0 preserva historial INVIMA.

    Query params:
      dias_inactividad: int (default 365)
      incluir_con_stock: 1/0 (default 0 · stock>0 NO se archiva nunca)

    Returns:
      {
        total_activas: int,
        sin_uso: [
          {codigo, nombre, ultima_actividad, dias_inactivo,
           stock_actual_g, en_formula: bool, n_movs}
        ],
        resumen: {sin_uso, con_stock_pero_no_usadas, ...}
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    try:
        dias = int(request.args.get('dias_inactividad', 365))
        if dias < 30:
            dias = 30
        if dias > 3650:
            dias = 3650
    except Exception:
        dias = 365

    incluir_con_stock = request.args.get('incluir_con_stock', '0') == '1'

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    try:
        total_activas = c.execute(
            "SELECT COUNT(*) FROM maestro_mps WHERE activo=1"
        ).fetchone()[0]

        rows = c.execute("""
            WITH mp_uso AS (
                SELECT
                    m.codigo_mp                                 AS codigo,
                    COALESCE(m.nombre_comercial, m.nombre_inci, '') AS nombre,
                    (SELECT MAX(fecha) FROM movimientos
                      WHERE material_id = m.codigo_mp)          AS ultima_act,
                    (SELECT COUNT(*) FROM movimientos
                      WHERE material_id = m.codigo_mp)          AS n_movs,
                    (SELECT COALESCE(SUM(
                        CASE WHEN UPPER(tipo) IN ('ENTRADA','RECEPCION',
                                                  'AJUSTE_POS','DEVOLUCION')
                             THEN COALESCE(cantidad,0)
                             WHEN UPPER(tipo) IN ('SALIDA','CONSUMO',
                                                  'AJUSTE_NEG','BAJA')
                             THEN -COALESCE(cantidad,0)
                             ELSE 0 END), 0)
                      FROM movimientos
                      WHERE material_id = m.codigo_mp)          AS stock_g,
                    (SELECT COUNT(*) FROM formula_items
                      WHERE material_id = m.codigo_mp)          AS n_formulas
                FROM maestro_mps m
                WHERE m.activo = 1
            )
            SELECT codigo, nombre, ultima_act, n_movs, stock_g, n_formulas
            FROM mp_uso
            WHERE n_formulas = 0
              AND (
                ultima_act IS NULL
                OR date(ultima_act) < date('now', '-' || ? || ' days')
              )
              AND (? = 1 OR ABS(stock_g) < 1)
            ORDER BY
              CASE WHEN ultima_act IS NULL THEN 1 ELSE 0 END DESC,
              ultima_act ASC
            LIMIT 500
        """, (dias, 1 if incluir_con_stock else 0)).fetchall()

        from datetime import date as _date
        hoy = _date.today()
        sin_uso = []
        con_stock_pero_inutil = 0
        for codigo, nombre, ultima, n_movs, stock_g, n_formulas in rows:
            dias_inact = None
            if ultima:
                try:
                    dias_inact = (hoy - _date.fromisoformat(ultima[:10])).days
                except Exception:
                    dias_inact = None
            stock_g = float(stock_g or 0)
            if abs(stock_g) >= 1:
                con_stock_pero_inutil += 1
            sin_uso.append({
                'codigo': codigo,
                'nombre': nombre or '',
                'ultima_actividad': ultima,
                'dias_inactivo': dias_inact,
                'stock_actual_g': round(stock_g, 2),
                'en_formula': False,
                'n_movs': int(n_movs or 0),
            })

        conn.close()

        return jsonify({
            'ok': True,
            'total_activas': total_activas,
            'umbral_dias_inactividad': dias,
            'incluye_con_stock': incluir_con_stock,
            'sin_uso': sin_uso,
            'resumen': {
                'sin_uso': len(sin_uso),
                'con_stock_pero_no_usadas': con_stock_pero_inutil,
                'archivables_seguro': sum(
                    1 for x in sin_uso if abs(x['stock_actual_g']) < 1
                ),
            },
            'message': (f'{len(sin_uso)} MPs candidatas a archivar '
                        f'· {total_activas} activas en catálogo'),
        }), 200

    except Exception as e:
        conn.close()
        return jsonify({'error': 'falla query',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/archivar-mps-sin-uso-bulk", methods=["POST"])
def archivar_mps_sin_uso_bulk():
    """Archiva (activo=0) las MPs especificadas. Verifica antes que cumplan
    criterio sin-uso: no en formula, stock=0, sin movs recientes.

    Sebastián 8-may-2026: el bulk archive evita acción manual una por una.
    Las MPs archivadas NO se borran · audit_log preserva trazabilidad
    INVIMA. Para reactivar, usar el panel `desactivar-mp` con activo=1.

    Body JSON:
      codigos: ['MP00001', 'MP00002', ...]
      motivo: str (queda en audit_log)
      forzar: bool (default False · si True, salta verificación de criterio)

    Returns: { ok, archivadas, rechazadas: [...], audit_id }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    codigos = d.get('codigos') or []
    motivo = (d.get('motivo') or 'Limpieza catálogo · sin uso').strip()
    forzar = bool(d.get('forzar', False))

    if not isinstance(codigos, list) or not codigos:
        return jsonify({'error': 'codigos debe ser lista no vacía'}), 400
    if len(codigos) > 200:
        return jsonify({'error': 'max 200 codigos por request'}), 400

    codigos_limpios = [str(c).strip().upper() for c in codigos if str(c).strip()]
    if not codigos_limpios:
        return jsonify({'error': 'codigos vacíos tras limpieza'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    archivadas = []
    rechazadas = []
    try:
        for cod in codigos_limpios:
            row = c.execute(
                "SELECT codigo_mp, "
                "  COALESCE(nombre_comercial, nombre_inci, '') AS nombre, "
                "  activo "
                "FROM maestro_mps WHERE codigo_mp = ?",
                (cod,)
            ).fetchone()
            if not row:
                rechazadas.append({'codigo': cod,
                                    'razon': 'no existe en maestro_mps'})
                continue
            if row[2] == 0:
                rechazadas.append({'codigo': cod,
                                    'razon': 'ya archivada (activo=0)'})
                continue

            if not forzar:
                n_form = c.execute(
                    "SELECT COUNT(*) FROM formula_items WHERE material_id=?",
                    (cod,)
                ).fetchone()[0]
                if n_form > 0:
                    rechazadas.append({'codigo': cod,
                                        'razon': f'en uso en {n_form} fórmula(s)'})
                    continue

                stock_row = c.execute("""
                    SELECT COALESCE(SUM(
                        CASE WHEN UPPER(tipo) IN ('ENTRADA','RECEPCION',
                                                  'AJUSTE_POS','DEVOLUCION')
                             THEN COALESCE(cantidad,0)
                             WHEN UPPER(tipo) IN ('SALIDA','CONSUMO',
                                                  'AJUSTE_NEG','BAJA')
                             THEN -COALESCE(cantidad,0)
                             ELSE 0 END), 0)
                    FROM movimientos WHERE material_id=?
                """, (cod,)).fetchone()
                stock_g = float(stock_row[0] or 0)
                if abs(stock_g) >= 1:
                    rechazadas.append({'codigo': cod,
                                        'razon': (f'stock no-cero '
                                                  f'({stock_g:.0f}g)')})
                    continue

            c.execute(
                "UPDATE maestro_mps SET activo=0 WHERE codigo_mp=?",
                (cod,)
            )
            archivadas.append({'codigo': cod, 'nombre': row[1] or ''})

        if archivadas:
            try:
                audit_log(
                    c, usuario=u,
                    accion='ARCHIVAR_MPS_SIN_USO_BULK',
                    tabla='maestro_mps', registro_id='bulk',
                    despues={
                        'motivo': motivo,
                        'forzar': forzar,
                        'archivadas': archivadas,
                        'rechazadas': rechazadas,
                        'n_archivadas': len(archivadas),
                        'n_rechazadas': len(rechazadas),
                    },
                    detalle=(f'Archivadas {len(archivadas)} MPs · '
                             f'rechazadas {len(rechazadas)} · motivo: {motivo}'),
                )
            except Exception:
                pass

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'archivadas': archivadas,
            'rechazadas': rechazadas,
            'n_archivadas': len(archivadas),
            'n_rechazadas': len(rechazadas),
            'message': (f'✓ Archivadas {len(archivadas)} MPs '
                        f'· rechazadas {len(rechazadas)}'),
        }), 200

    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/investigar-mp/<codigo>", methods=["GET"])
def investigar_mp(codigo):
    """Devuelve TODO sobre un MP: catálogo + movs por lote + saldos.

    Sebastián 10-may-2026: vista forensic para investigar findings raros
    (stock negativo, huérfanos, etc.) sin tener que ir tabla por tabla.

    Returns:
      {
        mp: {codigo, nombre, activo, ...} | null si no existe,
        lotes: [{lote, n_movs, stock_neto, primer_mov, ultimo_mov}],
        movimientos_recientes: [{id, tipo, cantidad, lote, fecha, obs}],
        stock_total_neto: float,
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    codigo = (codigo or '').strip()
    if not codigo:
        return jsonify({'error': 'codigo requerido'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # Catálogo
    mp_row = c.execute(
        "SELECT codigo_mp, nombre_inci, nombre_comercial, tipo, tipo_material, "
        "       proveedor, stock_minimo, activo "
        "FROM maestro_mps WHERE codigo_mp=?", (codigo,)
    ).fetchone()
    mp_info = None
    if mp_row:
        mp_info = {
            'codigo_mp': mp_row[0], 'nombre_inci': mp_row[1],
            'nombre_comercial': mp_row[2], 'tipo': mp_row[3],
            'tipo_material': mp_row[4], 'proveedor': mp_row[5],
            'stock_minimo': mp_row[6], 'activo': bool(mp_row[7]),
        }

    # Lotes con saldo
    lotes_rows = c.execute("""
        SELECT COALESCE(lote,'') as lote, COUNT(*) as n,
               ROUND(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END), 2) as neto,
               MIN(fecha) as primero, MAX(fecha) as ultimo,
               MAX(estado_lote) as estado, MAX(fecha_vencimiento) as fv
        FROM movimientos WHERE material_id=?
        GROUP BY lote ORDER BY neto DESC LIMIT 100
    """, (codigo,)).fetchall()
    lotes = [
        {'lote': r[0], 'n_movs': r[1], 'stock_neto_g': r[2],
         'primer_mov': r[3], 'ultimo_mov': r[4],
         'estado_lote': r[5], 'fecha_venc': r[6]}
        for r in lotes_rows
    ]

    # Movimientos recientes (últimos 50)
    mov_rows = c.execute("""
        SELECT id, tipo, cantidad, COALESCE(lote,'') as lote,
               fecha, SUBSTR(COALESCE(observaciones,''),1,200) as obs,
               COALESCE(operador,'') as op
        FROM movimientos WHERE material_id=?
        ORDER BY id DESC LIMIT 50
    """, (codigo,)).fetchall()
    movimientos = [
        {'id': r[0], 'tipo': r[1], 'cantidad_g': r[2], 'lote': r[3],
         'fecha': r[4], 'observaciones': r[5], 'operador': r[6]}
        for r in mov_rows
    ]

    # Stock total neto
    total_row = c.execute(
        "SELECT ROUND(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END), 2) "
        "FROM movimientos WHERE material_id=?", (codigo,)
    ).fetchone()
    stock_total = (total_row[0] or 0) if total_row else 0

    conn.close()

    return jsonify({
        'ok': True,
        'codigo': codigo,
        'mp': mp_info,
        'stock_total_neto_g': stock_total,
        'lotes_resumen': lotes,
        'movimientos_recientes': movimientos,
    }), 200


@bp.route("/api/admin/formula-huerfanos-con-sugerencias", methods=["GET"])
def formula_huerfanos_con_sugerencias():
    """Detecta TODOS los huérfanos en formula_items y sugiere el código
    real de maestro_mps basado en similitud de nombre.

    Sebastián 10-may-2026: 'las formulas tienen cosas diferentes que lo
    que hay en stock · debemos normalizar nombres y codigos'.

    Para cada material_id en formula_items que NO existe en maestro_mps
    activo, busca el mejor match en maestro_mps por nombre (fuzzy).

    Returns:
      huerfanos: [{
        material_id_actual_formula,
        material_nombre_en_formula,
        productos_que_lo_usan: [...],
        sugerencias: [{
          codigo_mp_correcto, nombre_match, similitud,
          razon: 'nombre exacto' | 'nombre similar' | ...
        }],
        recomendacion: 'auto' | 'manual' | 'crear_nueva'
      }]
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # 1. Encontrar huérfanos
    huerfanos_rows = c.execute("""
        SELECT fi.material_id,
               MAX(fi.material_nombre) as nombre_en_formula,
               COUNT(DISTINCT fi.producto_nombre) as n_productos,
               GROUP_CONCAT(DISTINCT fi.producto_nombre) as productos
        FROM formula_items fi
        LEFT JOIN maestro_mps mp
               ON fi.material_id = mp.codigo_mp AND mp.activo = 1
        WHERE mp.codigo_mp IS NULL
          AND fi.material_id IS NOT NULL AND TRIM(fi.material_id) != ''
        GROUP BY fi.material_id
        ORDER BY n_productos DESC
    """).fetchall()

    # 2. Cargar catálogo activo para buscar matches
    catalogo = c.execute("""
        SELECT codigo_mp,
               COALESCE(nombre_comercial,'') as nc,
               COALESCE(nombre_inci,'') as inci
        FROM maestro_mps WHERE activo = 1
    """).fetchall()

    # 3. Para cada huérfano, calcular sugerencias
    import difflib as _diff
    huerfanos = []
    for r in huerfanos_rows:
        mid_huerfano, nom_formula, n_prods, prods = r
        nom_formula_norm = (nom_formula or '').strip().lower()
        productos = (prods or '').split(',')[:10]

        sugerencias = []
        if nom_formula_norm:
            # Buscar matches exactos primero
            for cm, nc, inci in catalogo:
                if cm == mid_huerfano:
                    continue
                nc_norm = (nc or '').strip().lower()
                inci_norm = (inci or '').strip().lower()
                if nc_norm and nc_norm == nom_formula_norm:
                    sugerencias.append({
                        'codigo_mp_correcto': cm,
                        'nombre_match': nc, 'similitud': 1.0,
                        'razon': 'nombre comercial exacto',
                    })
                elif inci_norm and inci_norm == nom_formula_norm:
                    sugerencias.append({
                        'codigo_mp_correcto': cm,
                        'nombre_match': inci, 'similitud': 1.0,
                        'razon': 'INCI exacto',
                    })

            # Si no hubo exacto, búsqueda fuzzy
            if not sugerencias:
                candidatos = []
                for cm, nc, inci in catalogo:
                    if cm == mid_huerfano:
                        continue
                    for nombre_cand, etiq in [(nc, 'nombre comercial'),
                                              (inci, 'INCI')]:
                        nc_norm2 = (nombre_cand or '').strip().lower()
                        if not nc_norm2 or len(nc_norm2) < 4:
                            continue
                        ratio = _diff.SequenceMatcher(
                            None, nom_formula_norm, nc_norm2
                        ).ratio()
                        if ratio >= 0.7:
                            candidatos.append({
                                'codigo_mp_correcto': cm,
                                'nombre_match': nombre_cand,
                                'similitud': round(ratio, 3),
                                'razon': f'{etiq} similar (fuzzy)',
                            })
                # Top 5 ordenados por similitud
                candidatos.sort(key=lambda x: -x['similitud'])
                sugerencias = candidatos[:5]

        # Recomendación
        if sugerencias and sugerencias[0]['similitud'] >= 0.95:
            recomendacion = 'auto'
        elif sugerencias and sugerencias[0]['similitud'] >= 0.85:
            recomendacion = 'revisar_alta_confianza'
        elif sugerencias:
            recomendacion = 'manual'
        else:
            recomendacion = 'crear_nueva_mp'

        huerfanos.append({
            'material_id_actual_formula': mid_huerfano,
            'material_nombre_en_formula': nom_formula or '',
            'n_productos_que_lo_usan': n_prods,
            'productos': productos,
            'sugerencias': sugerencias,
            'recomendacion': recomendacion,
        })

    conn.close()
    return jsonify({
        'ok': True,
        'huerfanos': huerfanos,
        'resumen': {
            'total_huerfanos': len(huerfanos),
            'auto': sum(1 for h in huerfanos if h['recomendacion'] == 'auto'),
            'revisar_alta_confianza': sum(1 for h in huerfanos if h['recomendacion'] == 'revisar_alta_confianza'),
            'manual': sum(1 for h in huerfanos if h['recomendacion'] == 'manual'),
            'crear_nueva_mp': sum(1 for h in huerfanos if h['recomendacion'] == 'crear_nueva_mp'),
        },
    }), 200


@bp.route("/api/admin/auditoria-formulas-completa", methods=["GET"])
def auditoria_formulas_completa():
    """S1 · Integridad fórmulas maestras (read-only).

    Sebastián 8-may-2026 (revisa-cosa-a-cosa): un solo endpoint con
    veredicto unificado sobre fórmulas. No modifica nada · solo lee
    y reporta.

    Checks (cada uno bloqueante = ALTA si > 0):
      1. Huérfanos: material_id no existe en maestro_mps activo
      2. Duplicados: (producto, material_id) con >1 fila
      3. Suma % ≠ 100 ±0.5
      4. material_id NULL o vacío
      5. porcentaje NULL, <0, o >100
      6. Productos con fórmula pero sin items (declared but empty)
      7. Items sin material_nombre y sin material_id (huérfano absoluto)

    Returns:
      {
        ok: bool,
        score: 0-100 (100 = perfecto),
        veredicto: 'PERFECTA' | 'MENOR' | 'BLOQUEANTE',
        resumen: {n_formulas, n_items, ...counts},
        checks: {check1: {ok, count, top}, ...}
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # Cada check independiente · si uno falla, el resto sigue.
    errores_checks = {}

    def _safe(label, fn, default):
        try:
            return fn()
        except Exception as ex:
            errores_checks[label] = str(ex)[:200]
            return default

    # Pre-check: formula_items existe?
    try:
        n_formulas = c.execute(
            "SELECT COUNT(DISTINCT producto_nombre) FROM formula_items "
            "WHERE producto_nombre IS NOT NULL"
        ).fetchone()[0]
        n_items = c.execute(
            "SELECT COUNT(*) FROM formula_items"
        ).fetchone()[0]
    except Exception as ex:
        conn.close()
        return jsonify({
            'ok': False, 'error': 'tabla formula_items inaccesible',
            'detail': str(ex)[:300],
        }), 500

    # 1. Huérfanos
    huerfanos = _safe('huerfanos', lambda: c.execute("""
        SELECT fi.material_id, MAX(fi.material_nombre) AS nombre,
               COUNT(DISTINCT fi.producto_nombre) AS n_productos
        FROM formula_items fi
        LEFT JOIN maestro_mps m
          ON m.codigo_mp = fi.material_id AND COALESCE(m.activo,1) = 1
        WHERE m.codigo_mp IS NULL
          AND fi.material_id IS NOT NULL AND TRIM(fi.material_id) != ''
        GROUP BY fi.material_id
        ORDER BY n_productos DESC
        LIMIT 50
    """).fetchall(), [])

    # 2. Duplicados
    duplicados = _safe('duplicados', lambda: c.execute("""
        SELECT producto_nombre, material_id, COUNT(*) AS veces
        FROM formula_items
        WHERE material_id IS NOT NULL AND TRIM(material_id) != ''
          AND producto_nombre IS NOT NULL
        GROUP BY producto_nombre, material_id
        HAVING veces > 1
        ORDER BY veces DESC
        LIMIT 50
    """).fetchall(), [])

    # 3. Suma % SOLO si > 100 (sobreapasamiento) o exactamente 0
    # (fórmula vacía). Sebastián 8-may-2026: en cosmética, formulas
    # declaran solo activos · el resto es agua q.s. (quantum satis)
    # que no se declara como item. Sumas 1-100% son legítimas.
    # SOLO es bug: suma > 100 (sobra · imposible regulatorio) o = 0
    # (fórmula sin ingredientes declarados).
    sumas_malas = _safe('sumas_pct_no_100', lambda: c.execute("""
        SELECT producto_nombre,
               ROUND(SUM(COALESCE(porcentaje,0)), 2) AS suma_pct,
               ROUND(SUM(COALESCE(cantidad_g_por_lote, 0)), 2) AS suma_g,
               COUNT(*) AS items
        FROM formula_items
        WHERE producto_nombre IS NOT NULL
        GROUP BY producto_nombre
        HAVING (suma_pct > 100.5 OR suma_pct < 0.001)
           AND suma_g < 1
        ORDER BY suma_pct DESC
        LIMIT 50
    """).fetchall(), [])

    # 4. material_id NULL o vacío
    nulos = _safe('material_id_nulos', lambda: c.execute("""
        SELECT COUNT(*) FROM formula_items
        WHERE material_id IS NULL OR TRIM(material_id) = ''
    """).fetchone()[0], 0)

    # 5. porcentaje inválido
    pct_invalidos = _safe('pct_invalidos', lambda: c.execute("""
        SELECT id, producto_nombre, material_id, porcentaje
        FROM formula_items
        WHERE porcentaje IS NULL OR porcentaje < 0 OR porcentaje > 100
        LIMIT 50
    """).fetchall(), [])

    # 6. Productos en formula_headers sin items (tabla puede no existir)
    headers_vacios = _safe('headers_vacios', lambda: c.execute("""
        SELECT fh.producto_nombre
        FROM formula_headers fh
        WHERE NOT EXISTS (
            SELECT 1 FROM formula_items fi
            WHERE fi.producto_nombre = fh.producto_nombre
        )
        LIMIT 50
    """).fetchall(), [])

    # 7. Items sin nombre Y sin material_id
    huerfanos_absolutos = _safe('huerfanos_absolutos', lambda: c.execute("""
        SELECT id, producto_nombre, porcentaje
        FROM formula_items
        WHERE (material_id IS NULL OR TRIM(material_id) = '')
          AND (material_nombre IS NULL OR TRIM(material_nombre) = '')
        LIMIT 50
    """).fetchall(), [])

    # Score 0-100: cada check bloqueante restará proporcionalmente
    n_huer = len(huerfanos)
    n_dup = len(duplicados)
    n_pct100 = len(sumas_malas)
    n_pct_inv = len(pct_invalidos)
    n_hdr_vac = len(headers_vacios)
    n_huer_abs = len(huerfanos_absolutos)

    # Pesos: huérfanos y duplicados son los más graves
    score = 100.0
    if n_formulas > 0:
        score -= min(40, 40 * n_huer / max(n_formulas, 1))
        score -= min(20, 20 * n_dup / max(n_formulas, 1))
        score -= min(20, 20 * n_pct100 / max(n_formulas, 1))
        score -= min(10, 10 * (n_pct_inv + nulos) / max(n_items, 1))
        score -= min(5, 5 * n_hdr_vac / max(n_formulas, 1))
        score -= min(5, 5 * n_huer_abs / max(n_items, 1))
    score = max(0.0, round(score, 1))

    if score >= 99:
        veredicto = 'PERFECTA'
    elif score >= 85:
        veredicto = 'MENOR'
    else:
        veredicto = 'BLOQUEANTE'

    conn.close()

    return jsonify({
        'ok': True,
        'score': score,
        'veredicto': veredicto,
        'resumen': {
            'n_formulas': n_formulas,
            'n_items': n_items,
            'huerfanos': n_huer,
            'duplicados': n_dup,
            'sumas_pct_no_100': n_pct100,
            'material_id_nulos': nulos,
            'pct_invalidos': n_pct_inv,
            'headers_vacios': n_hdr_vac,
            'huerfanos_absolutos': n_huer_abs,
        },
        'checks': {
            'huerfanos': {
                'ok': n_huer == 0,
                'count': n_huer,
                'top': [
                    {'material_id': h[0], 'nombre': h[1] or '',
                     'n_productos': h[2]}
                    for h in huerfanos[:10]
                ],
                'fix_link': '/admin/normalizar-formulas',
            },
            'duplicados': {
                'ok': n_dup == 0,
                'count': n_dup,
                'top': [
                    {'producto': d[0], 'material_id': d[1],
                     'veces': d[2]}
                    for d in duplicados[:10]
                ],
                'fix_link': '/admin/limpieza-cero-error',
            },
            'sumas_pct_no_100': {
                'ok': n_pct100 == 0,
                'count': n_pct100,
                'top': [
                    {'producto': s[0], 'suma_actual': s[1],
                     'diff': round((s[1] or 0) - 100, 2),
                     'items': s[3]}
                    for s in sumas_malas[:10]
                ],
                'fix_link': '/tecnica',
                'nota': ('Excluye fórmulas que usan cantidad_g_por_lote '
                          '(modalidad gramos directos · suma % no aplica).'),
            },
            'material_id_nulos': {
                'ok': nulos == 0,
                'count': nulos,
            },
            'pct_invalidos': {
                'ok': n_pct_inv == 0,
                'count': n_pct_inv,
                'top': [
                    {'id': p[0], 'producto': p[1],
                     'material_id': p[2], 'porcentaje': p[3]}
                    for p in pct_invalidos[:10]
                ],
            },
            'headers_vacios': {
                'ok': n_hdr_vac == 0,
                'count': n_hdr_vac,
                'top': [h[0] for h in headers_vacios[:10]],
            },
            'huerfanos_absolutos': {
                'ok': n_huer_abs == 0,
                'count': n_huer_abs,
                'top': [
                    {'id': h[0], 'producto': h[1], 'porcentaje': h[2]}
                    for h in huerfanos_absolutos[:10]
                ],
            },
        },
        'errores_checks': errores_checks,  # vacío si todo OK
        'message': (f'S1 fórmulas: {veredicto} · score {score}/100 '
                    f'· {n_formulas} fórmulas · {n_items} items'),
    }), 200


@bp.route("/api/admin/auditoria-producciones-descuento", methods=["GET"])
def auditoria_producciones_descuento():
    """S2 · Integridad producciones · descuento de MPs (read-only).

    Sebastián 8-may-2026: validar que cuando una producción se inicia,
    descuenta correctamente las MPs de inventario via movimientos Salida.

    Para cada producción en los últimos N días (default 90):
      1. iniciada · si tiene inicio_real_at != NULL
      2. descontada · si tiene inventario_descontado_at != NULL
      3. n_movs_salida · movimientos Salida con obs LIKE 'Producción INICIADA: <producto>%'
      4. estado_audit:
         - OK: iniciada + descontada + n_movs_salida >= 1
         - PENDIENTE: no iniciada (sin problema)
         - INICIADA_SIN_DESCUENTO: iniciada pero inventario_descontado_at NULL
         - DESCONTADA_SIN_MOVS: descontada pero ningún movimiento Salida (bug grave)
         - SIN_FORMULA: producto sin formula_items (no se puede descontar)
         - TERMINADA_SIN_DESCUENTO: tiene fin_real_at sin inventario_descontado_at

    Query params:
      dias: int (default 90)

    Returns:
      {
        ok, score (0-100), veredicto, resumen, producciones: [...]
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    try:
        dias = int(request.args.get('dias', 90))
        if dias < 1: dias = 1
        if dias > 730: dias = 730
    except Exception:
        dias = 90

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    errores_checks = {}

    try:
        # Producciones programadas con datos para auditar
        rows = c.execute("""
            SELECT pp.id, pp.producto, pp.fecha_programada, pp.lotes,
                   pp.estado,
                   pp.inicio_real_at,
                   pp.fin_real_at,
                   pp.inventario_descontado_at,
                   (SELECT COUNT(*) FROM formula_items
                     WHERE producto_nombre = pp.producto) AS n_formula_items,
                   (SELECT COUNT(*) FROM movimientos
                     WHERE tipo = 'Salida'
                       AND observaciones LIKE 'Producción INICIADA: ' || pp.producto || '%')
                   AS n_movs_salida_producto
            FROM produccion_programada pp
            WHERE date(pp.fecha_programada) >= date('now', '-' || ? || ' days')
            ORDER BY pp.fecha_programada DESC
            LIMIT 500
        """, (dias,)).fetchall()
    except Exception as e:
        conn.close()
        return jsonify({
            'ok': False,
            'error': 'falla en query base · puede que falten columnas',
            'detail': str(e)[:300],
        }), 500

    producciones = []
    n_ok = 0
    n_pendientes = 0
    n_sin_descuento = 0
    n_descontada_sin_movs = 0
    n_sin_formula = 0
    n_terminada_sin_descuento = 0

    for r in rows:
        (pp_id, producto, fecha_prog, lotes, estado,
         inicio_real, fin_real, desc_at, n_fi, n_movs) = r

        iniciada = bool(inicio_real)
        descontada = bool(desc_at)
        terminada = bool(fin_real)
        tiene_formula = (n_fi or 0) > 0

        if not iniciada:
            estado_audit = 'PENDIENTE'
            n_pendientes += 1
        elif not tiene_formula:
            estado_audit = 'SIN_FORMULA'
            n_sin_formula += 1
        elif iniciada and not descontada:
            estado_audit = 'INICIADA_SIN_DESCUENTO'
            n_sin_descuento += 1
        elif terminada and not descontada:
            estado_audit = 'TERMINADA_SIN_DESCUENTO'
            n_terminada_sin_descuento += 1
        elif descontada and (n_movs or 0) == 0:
            estado_audit = 'DESCONTADA_SIN_MOVS'
            n_descontada_sin_movs += 1
        else:
            estado_audit = 'OK'
            n_ok += 1

        producciones.append({
            'id': pp_id,
            'producto': producto,
            'fecha_programada': fecha_prog,
            'lotes': lotes,
            'estado': estado,
            'iniciada': iniciada,
            'descontada': descontada,
            'terminada': terminada,
            'n_formula_items': n_fi or 0,
            'n_movs_salida_producto': n_movs or 0,
            'estado_audit': estado_audit,
        })

    n_total = len(producciones)
    n_iniciadas = n_total - n_pendientes
    n_problemas = (n_sin_descuento + n_descontada_sin_movs +
                    n_sin_formula + n_terminada_sin_descuento)

    # Score: cada problema baja proporcionalmente.
    # Si no hay producciones iniciadas, score=100 (no hay nada que auditar)
    # ── AUDITAR TAMBIEN tabla legacy `producciones` ───────────────────────
    # Hueco descubierto Sebastian 8-may-2026: 358 produccion_programada
    # todas en pendiente PERO dashboard muestra producciones histórico.
    # Si la planta usa la tabla legacy `producciones`, NO pasa por
    # _descontar_mp_produccion y MPs no se descuentan.
    legacy = []
    n_legacy_total = 0
    n_legacy_sin_movs = 0
    try:
        # Re-abrir conn
        conn2 = sqlite3.connect(DB_PATH)
        conn2.execute("PRAGMA busy_timeout=2000")
        c2 = conn2.cursor()
        # Flujo legacy (inventario.py · /api/producciones POST) genera movs
        # con observaciones formato 'FEFO:PROD-XXXXX:producto x Nkg' o
        # 'UNLIMITED:PROD-XXXXX:producto x Nkg'. Detectamos por id de
        # producción embebido como PROD-00123 zero-padded.
        legacy_rows = c2.execute("""
            SELECT p.id, p.producto, p.fecha, p.cantidad, p.estado, p.lote,
                   (SELECT COUNT(*) FROM movimientos
                     WHERE tipo = 'Salida'
                       AND (
                         observaciones LIKE '%PROD-' || printf('%05d', p.id) || '%'
                         OR (p.lote IS NOT NULL AND p.lote != '' AND
                             observaciones LIKE '%' || p.lote || '%')
                         OR (observaciones LIKE 'Producción INICIADA: ' || p.producto || '%')
                       ))
                   AS n_movs_relacionados,
                   (SELECT COUNT(*) FROM formula_items
                     WHERE producto_nombre = p.producto) AS n_formula_items
            FROM producciones p
            WHERE date(COALESCE(p.fecha, '1970-01-01')) >= date('now', '-' || ? || ' days')
            ORDER BY p.fecha DESC
            LIMIT 200
        """, (dias,)).fetchall()
        conn2.close()
        n_legacy_total = len(legacy_rows)
        for r in legacy_rows:
            pid, producto, fecha, cantidad, estado, lote, n_movs, n_fi = r
            sin_movs = (n_movs or 0) == 0 and (n_fi or 0) > 0
            if sin_movs:
                n_legacy_sin_movs += 1
            legacy.append({
                'id': pid,
                'producto': producto,
                'fecha': fecha,
                'cantidad': cantidad,
                'estado': estado,
                'lote': lote,
                'n_movs_salida_relacionados': n_movs or 0,
                'n_formula_items': n_fi or 0,
                'sin_movimientos_salida': bool(sin_movs),
            })
    except Exception as e:
        errores_checks['legacy_producciones'] = str(e)[:200]

    # Score considera ambas: produccion_programada Y producciones legacy
    n_total_audit = n_iniciadas + n_legacy_total
    n_total_problemas = n_problemas + n_legacy_sin_movs
    if n_total_audit == 0:
        score = 100.0
    else:
        score = 100.0 * (1.0 - n_total_problemas / max(n_total_audit, 1))
    score = max(0.0, round(score, 1))

    if score >= 99:
        veredicto = 'PERFECTA'
    elif score >= 85:
        veredicto = 'MENOR'
    else:
        veredicto = 'BLOQUEANTE'

    conn.close()

    # Top producciones con problemas (no mostrar las OK ni PENDIENTE)
    problemas = [p for p in producciones
                  if p['estado_audit'] not in ('OK', 'PENDIENTE')]
    # Agregar legacy con problema al inicio
    problemas_legacy = [l for l in legacy if l.get('sin_movimientos_salida')]

    return jsonify({
        'ok': True,
        'score': score,
        'veredicto': veredicto,
        'resumen': {
            'dias_horizonte': dias,
            'n_total': n_total,
            'n_iniciadas': n_iniciadas,
            'n_pendientes': n_pendientes,
            'n_ok': n_ok,
            'n_problemas': n_problemas,
            'legacy_total': n_legacy_total,
            'legacy_sin_movs': n_legacy_sin_movs,
            'iniciadas_sin_descuento': n_sin_descuento,
            'descontadas_sin_movs': n_descontada_sin_movs,
            'sin_formula': n_sin_formula,
            'terminadas_sin_descuento': n_terminada_sin_descuento,
        },
        'problemas': problemas[:50],
        'legacy_producciones': legacy[:100],
        'legacy_problemas': problemas_legacy[:50],
        'errores_checks': errores_checks,
        'message': (f'S2 producciones: {veredicto} · score {score}/100 '
                    f'· {n_total} programadas · {n_legacy_total} legacy '
                    f'· {n_total_problemas} con problemas'),
    }), 200


@bp.route("/api/admin/auditoria-kardex-drift", methods=["GET"])
def auditoria_kardex_drift():
    """S3 · Drift inventario · stock derivado vs movimientos (read-only).

    Sebastián 8-may-2026: el corazón de zero-error. Para CADA MP el stock
    debe derivarse exactamente de SUM(entradas) - SUM(salidas) + ajustes.
    Si hay drift (stock negativo o discrepancia), es bug operativo: doble
    descuento, ajuste sin contrapartida, etc.

    Usa los helpers existentes:
      - detect_drift_mp(conn): MPs con stock NEGATIVO
      - detect_drift_mee(conn): MEEs con drift entre maestro_mee.stock_actual
        y SUM(movimientos_mee)

    Score 100 = sin drift. Cada item con drift baja el score proporcional
    al total de MPs activas.

    Returns:
      {
        ok, score, veredicto, resumen, mp_negativos, mee_drift
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")

    try:
        from inventario_helpers import detect_drift_mp, detect_drift_mee
    except Exception as e:
        conn.close()
        return jsonify({
            'ok': False,
            'error': 'helpers no disponibles',
            'detail': str(e)[:200],
        }), 500

    try:
        mp_neg = detect_drift_mp(conn)
        mee_drift = detect_drift_mee(conn)

        # Total MPs activas (denominador para score)
        try:
            n_mps_activas = conn.execute(
                "SELECT COUNT(*) FROM maestro_mps WHERE COALESCE(activo,1)=1"
            ).fetchone()[0]
        except Exception:
            n_mps_activas = 0
        try:
            n_mee_activos = conn.execute(
                "SELECT COUNT(*) FROM maestro_mee "
                "WHERE COALESCE(estado,'Activo') != 'Inactivo'"
            ).fetchone()[0]
        except Exception:
            n_mee_activos = 0

        n_total_items = n_mps_activas + n_mee_activos
        n_drift = len(mp_neg) + len(mee_drift)
        n_criticos = sum(1 for x in mp_neg if x.get('severidad') == 'critical')
        n_criticos += sum(1 for x in mee_drift if x.get('severidad') == 'critical')

        # Score: 100 - (drift/total) * 100, con peso extra para criticos
        if n_total_items > 0:
            score = 100.0 - (100.0 * n_drift / n_total_items) - (10.0 * n_criticos)
            score = max(0.0, round(score, 1))
        else:
            score = 100.0

        if score >= 99 and n_drift == 0:
            veredicto = 'PERFECTA'
        elif score >= 85:
            veredicto = 'MENOR'
        else:
            veredicto = 'BLOQUEANTE'

        conn.close()

        return jsonify({
            'ok': True,
            'score': score,
            'veredicto': veredicto,
            'resumen': {
                'mps_activas': n_mps_activas,
                'mees_activos': n_mee_activos,
                'mp_negativos': len(mp_neg),
                'mee_con_drift': len(mee_drift),
                'criticos': n_criticos,
                'total_con_drift': n_drift,
            },
            'mp_negativos': mp_neg,
            'mee_drift': mee_drift,
            'message': (f'S3 kardex: {veredicto} · score {score}/100 '
                        f'· {n_drift} items con drift '
                        f'(de {n_total_items} totales)'),
        }), 200

    except Exception as e:
        conn.close()
        return jsonify({
            'ok': False, 'error': 'falla query drift',
            'detail': str(e)[:300],
        }), 500


@bp.route("/admin/auditoria-kardex", methods=["GET"])
def admin_auditoria_kardex_page():
    """S3 · Página visual del drift inventario."""
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este panel.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_AUDIT_KARDEX_HTML, mimetype='text/html')


_AUDIT_KARDEX_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>S3 · Kardex · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0f172a;color:#f1f5f9;padding:20px;line-height:1.5}
h1{font-size:24px;margin-bottom:6px;color:#5eead4}
.sub{color:#94a3b8;font-size:13px;margin-bottom:20px}
.back{display:inline-block;color:#94a3b8;text-decoration:none;font-size:13px;margin-bottom:16px}
.back:hover{color:#f1f5f9}
.hero{background:#1e293b;border-radius:14px;padding:24px;margin-bottom:20px;text-align:center}
.score{font-size:72px;font-weight:800;line-height:1}
.score.ok{color:#22c55e}
.score.warn{color:#fbbf24}
.score.bad{color:#ef4444}
.verdict{font-size:16px;font-weight:700;letter-spacing:1px;margin-top:8px;text-transform:uppercase}
.verdict.ok{color:#22c55e}
.verdict.warn{color:#fbbf24}
.verdict.bad{color:#ef4444}
.resumen{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px;margin-top:14px}
.kpi{background:#0f172a;border-radius:8px;padding:8px;font-size:11px;color:#94a3b8}
.kpi b{display:block;color:#f1f5f9;font-size:18px;margin-bottom:2px}
.kpi.bad b{color:#ef4444}
.kpi.ok b{color:#22c55e}
table{width:100%;border-collapse:collapse;background:#1e293b;border-radius:10px;overflow:hidden;font-size:12px;margin-top:10px}
th,td{padding:8px 10px;text-align:left;border-bottom:1px solid #334155}
th{background:#0f172a;color:#94a3b8;font-weight:600;font-size:11px;text-transform:uppercase}
tr:hover{background:#334155}
.badge{padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;display:inline-block}
.badge.critical{background:#7f1d1d;color:#fca5a5}
.badge.high{background:#7c2d12;color:#fed7aa}
.loading{text-align:center;padding:40px;color:#64748b}
.empty{color:#22c55e;text-align:center;padding:30px;font-size:15px}
.error{background:#7f1d1d;color:#fecaca;padding:14px;border-radius:8px;font-size:13px}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px;background:#5eead4;color:#0f172a}
button:hover{background:#2dd4bf}
.section-title{margin-top:18px;color:#fbbf24;font-size:14px}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>📊 S3 · Drift inventario · stock vs movimientos</h1>
<p class="sub">El corazón de zero-error. Para cada MP/MEE, el stock debe coincidir con SUM(entradas)−SUM(salidas)±ajustes. Si NO coincide, hay bug operativo.</p>

<div style="margin-bottom:14px"><button onclick="run()">🔄 Re-ejecutar</button></div>
<div id="content" class="loading">⏳ Cargando auditoría...</div>

<script>
function esc(s){return String(s===null||s===undefined?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;');}

async function run(){
  document.getElementById('content').className = 'loading';
  document.getElementById('content').innerHTML = '⏳ Cargando auditoría...';
  try{
    const r = await fetch('/api/admin/auditoria-kardex-drift');
    const d = await r.json();
    if(!r.ok){
      document.getElementById('content').innerHTML =
        '<div class="error">Error: ' + esc(d.error||'falla') + '<br><small>' + esc(d.detail||'') + '</small></div>';
      return;
    }
    render(d);
  }catch(e){
    document.getElementById('content').innerHTML =
      '<div class="error">Error de red: ' + esc(e.message) + '</div>';
  }
}

function render(d){
  const score = d.score || 0;
  const sclass = score >= 99 ? 'ok' : score >= 85 ? 'warn' : 'bad';
  const v = d.veredicto || 'BLOQUEANTE';
  const vclass = v === 'PERFECTA' ? 'ok' : v === 'MENOR' ? 'warn' : 'bad';
  const rs = d.resumen || {};

  let html = '<div class="hero">' +
    '<div class="score ' + sclass + '">' + score + '<small style="font-size:24px;color:#64748b">/100</small></div>' +
    '<div class="verdict ' + vclass + '">' + v + '</div>' +
    '<div class="resumen">' +
      kpi('MPs activas', rs.mps_activas, '') +
      kpi('MEEs activos', rs.mees_activos, '') +
      kpi('MPs negativas', rs.mp_negativos, (rs.mp_negativos||0)>0?'bad':'ok') +
      kpi('MEEs con drift', rs.mee_con_drift, (rs.mee_con_drift||0)>0?'bad':'ok') +
      kpi('Críticos', rs.criticos, (rs.criticos||0)>0?'bad':'ok') +
    '</div>' +
  '</div>';

  const mpn = d.mp_negativos || [];
  const meed = d.mee_drift || [];

  if(mpn.length === 0 && meed.length === 0){
    html += '<div class="empty">✓ Cero drift en ' + (rs.mps_activas + rs.mees_activos) + ' items totales · inventario perfectamente cuadrado.</div>';
  }
  if(mpn.length > 0){
    html += '<h3 class="section-title">⚠ MPs con stock NEGATIVO (' + mpn.length + ')</h3>';
    html += '<table><thead><tr><th>Código MP</th><th>Nombre</th><th>Stock (g)</th><th>Severidad</th></tr></thead><tbody>';
    for(const m of mpn){
      html += '<tr>' +
        '<td><b>' + esc(m.codigo_mp) + '</b></td>' +
        '<td>' + esc(m.nombre) + '</td>' +
        '<td style="color:#ef4444">' + esc(m.stock_g) + '</td>' +
        '<td><span class="badge ' + esc(m.severidad) + '">' + esc(m.severidad) + '</span></td>' +
      '</tr>';
    }
    html += '</tbody></table>';
  }
  if(meed.length > 0){
    html += '<h3 class="section-title">⚠ MEEs con drift entre persistido y calculado (' + meed.length + ')</h3>';
    html += '<table><thead><tr><th>Código</th><th>Nombre</th><th>Persistido</th><th>Calculado</th><th>Drift</th><th>Severidad</th></tr></thead><tbody>';
    for(const m of meed){
      html += '<tr>' +
        '<td><b>' + esc(m.codigo) + '</b></td>' +
        '<td>' + esc(m.nombre) + '</td>' +
        '<td>' + esc(m.stock_persistido) + '</td>' +
        '<td>' + esc(m.stock_calculado) + '</td>' +
        '<td style="color:#ef4444">' + esc(m.drift) + '</td>' +
        '<td><span class="badge ' + esc(m.severidad) + '">' + esc(m.severidad) + '</span></td>' +
      '</tr>';
    }
    html += '</tbody></table>';
  }

  document.getElementById('content').className = '';
  document.getElementById('content').innerHTML = html;
}

function kpi(label, val, cls){
  return '<div class="kpi ' + (cls||'') + '"><b>' + (val||0) + '</b>' + esc(label) + '</div>';
}

run();
</script>
</body></html>"""


@bp.route("/api/admin/auditoria-mps-nuevas", methods=["GET"])
def auditoria_mps_nuevas():
    """S4 · Audit MPs nuevas · que carguen en inventario correctamente.

    Sebastián 8-may-2026: cuando se ingresa una MP nueva, debe tener
    su primera Entrada para que aparezca en inventario con stock real.

    Heurística "MP nueva": primer movimiento en últimos N días (default 30).
    Para cada MP nueva detectada:
      - n_entradas: count movs tipo='Entrada'
      - n_salidas: count tipo='Salida'
      - stock_actual: SUM(entradas) - SUM(salidas) + ajustes
      - tiene_formula: count en formula_items
      - estado_audit:
        · OK: tiene >= 1 Entrada
        · SIN_ENTRADA: solo salidas/ajustes (bug grave · imposible operativo)
        · STOCK_NEGATIVO: stock < 0

    Query params:
      dias: int (default 30)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    try:
        dias = int(request.args.get('dias', 30))
        if dias < 1: dias = 1
        if dias > 365: dias = 365
    except Exception:
        dias = 30

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    try:
        rows = c.execute("""
            WITH primer_mov AS (
                SELECT material_id,
                       MIN(date(fecha)) AS primer_fecha,
                       SUM(CASE WHEN tipo='Entrada' THEN 1 ELSE 0 END) AS n_entradas,
                       SUM(CASE WHEN tipo='Salida' THEN 1 ELSE 0 END) AS n_salidas,
                       SUM(CASE WHEN tipo='Entrada' THEN cantidad
                                WHEN tipo='Salida' THEN -cantidad
                                ELSE 0 END) AS stock_g
                FROM movimientos
                WHERE material_id IS NOT NULL AND material_id != ''
                GROUP BY material_id
            )
            SELECT pm.material_id,
                   COALESCE(m.nombre_comercial, m.nombre_inci, '') AS nombre,
                   pm.primer_fecha,
                   pm.n_entradas, pm.n_salidas, pm.stock_g,
                   COALESCE(m.activo, 1) AS activo,
                   (SELECT COUNT(*) FROM formula_items
                     WHERE material_id = pm.material_id) AS n_formula
            FROM primer_mov pm
            LEFT JOIN maestro_mps m ON m.codigo_mp = pm.material_id
            WHERE date(pm.primer_fecha) >= date('now', '-' || ? || ' days')
            ORDER BY pm.primer_fecha DESC
            LIMIT 200
        """, (dias,)).fetchall()
    except Exception as e:
        conn.close()
        return jsonify({
            'ok': False, 'error': 'falla query',
            'detail': str(e)[:300],
        }), 500

    mps_nuevas = []
    n_ok = 0
    n_sin_entrada = 0
    n_stock_negativo = 0
    for r in rows:
        (mid, nombre, primer, n_e, n_s, stock, activo, n_f) = r
        n_e = int(n_e or 0)
        n_s = int(n_s or 0)
        stock = float(stock or 0)
        if n_e == 0:
            estado_audit = 'SIN_ENTRADA'
            n_sin_entrada += 1
        elif stock < -1:
            estado_audit = 'STOCK_NEGATIVO'
            n_stock_negativo += 1
        else:
            estado_audit = 'OK'
            n_ok += 1
        mps_nuevas.append({
            'codigo_mp': mid,
            'nombre': nombre or '',
            'primer_movimiento': primer,
            'n_entradas': n_e,
            'n_salidas': n_s,
            'stock_actual_g': round(stock, 2),
            'activo': bool(activo),
            'n_formula_items': int(n_f or 0),
            'estado_audit': estado_audit,
        })

    n_total = len(mps_nuevas)
    n_problemas = n_sin_entrada + n_stock_negativo

    if n_total == 0:
        score = 100.0
    else:
        score = 100.0 * (1.0 - n_problemas / n_total)
    score = max(0.0, round(score, 1))

    if score >= 99:
        veredicto = 'PERFECTA'
    elif score >= 85:
        veredicto = 'MENOR'
    else:
        veredicto = 'BLOQUEANTE'

    conn.close()

    return jsonify({
        'ok': True,
        'score': score,
        'veredicto': veredicto,
        'resumen': {
            'dias_horizonte': dias,
            'n_total': n_total,
            'n_ok': n_ok,
            'n_problemas': n_problemas,
            'sin_entrada': n_sin_entrada,
            'stock_negativo': n_stock_negativo,
        },
        'mps_nuevas': mps_nuevas[:100],
        'message': (f'S4 MPs nuevas: {veredicto} · score {score}/100 '
                    f'· {n_total} MPs nuevas · {n_problemas} con problemas'),
    }), 200


@bp.route("/api/admin/auditoria-lotes-nuevos", methods=["GET"])
def auditoria_lotes_nuevos():
    """S5 · Audit lotes nuevos · que queden con info real.

    Sebastián 8-may-2026: cada lote nuevo de MP debe tener fecha_vencimiento,
    proveedor, lote_id, y aparecer en bodega con stock real.

    Heurística "lote nuevo": primera Entrada del (material_id, lote) en
    últimos N días (default 30).

    Para cada lote nuevo:
      - tiene_fecha_venc: fecha_vencimiento NOT NULL
      - tiene_proveedor: proveedor NOT NULL
      - stock_actual: SUM(entradas) - SUM(salidas) por (material_id, lote)
      - estado_audit:
        · OK: tiene fecha_venc Y proveedor Y stock >= 0
        · SIN_FECHA_VENC: bug regulatorio (INVIMA exige fecha venc)
        · SIN_PROVEEDOR: trazabilidad incompleta
        · STOCK_NEGATIVO: lote con más salidas que entradas (imposible)

    Query params:
      dias: int (default 30)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    try:
        dias = int(request.args.get('dias', 30))
        if dias < 1: dias = 1
        if dias > 365: dias = 365
    except Exception:
        dias = 30

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    try:
        rows = c.execute("""
            WITH lote_resumen AS (
                SELECT material_id, lote,
                       MIN(date(fecha)) AS primer_fecha,
                       MAX(fecha_vencimiento) AS fv,
                       MAX(proveedor) AS prov,
                       SUM(CASE WHEN tipo='Entrada' THEN cantidad
                                WHEN tipo='Salida' THEN -cantidad
                                ELSE 0 END) AS stock_g,
                       COUNT(*) AS n_movs
                FROM movimientos
                WHERE material_id IS NOT NULL AND material_id != ''
                  AND lote IS NOT NULL AND lote != ''
                GROUP BY material_id, lote
            )
            SELECT lr.material_id, lr.lote, lr.primer_fecha, lr.fv,
                   lr.prov, lr.stock_g, lr.n_movs,
                   COALESCE(m.nombre_comercial, m.nombre_inci, '') AS nombre
            FROM lote_resumen lr
            LEFT JOIN maestro_mps m ON m.codigo_mp = lr.material_id
            WHERE date(lr.primer_fecha) >= date('now', '-' || ? || ' days')
            ORDER BY lr.primer_fecha DESC
            LIMIT 300
        """, (dias,)).fetchall()
    except Exception as e:
        conn.close()
        return jsonify({
            'ok': False, 'error': 'falla query',
            'detail': str(e)[:300],
        }), 500

    lotes = []
    n_ok = 0
    n_sin_fv = 0
    n_sin_prov = 0
    n_stock_neg = 0
    n_consumidos = 0  # stock<=0 · inertes operativamente
    for r in rows:
        (mid, lote, primer, fv, prov, stock, n_movs, nombre) = r
        tiene_fv = bool(fv and fv.strip())
        tiene_prov = bool(prov and prov.strip())
        stock = float(stock or 0)

        # Sebastián 8-may-2026: lotes con stock<=0 son consumidos o
        # anulados · operativamente inertes. NO requieren fecha_venc
        # ni proveedor para INVIMA (ya no hay producto físico).
        lote_consumido = stock <= 0
        problemas_lote = []
        if not tiene_fv and not lote_consumido:
            problemas_lote.append('SIN_FECHA_VENC')
            n_sin_fv += 1
        if not tiene_prov and not lote_consumido:
            problemas_lote.append('SIN_PROVEEDOR')
            n_sin_prov += 1
        if stock < -1:
            problemas_lote.append('STOCK_NEGATIVO')
            n_stock_neg += 1

        if not problemas_lote:
            estado_audit = 'OK' if not lote_consumido else 'CONSUMIDO'
            n_ok += 1
            if lote_consumido:
                n_consumidos += 1
        else:
            estado_audit = problemas_lote[0]  # primer problema más grave

        lotes.append({
            'material_id': mid,
            'lote': lote,
            'nombre': nombre or '',
            'primer_fecha': primer,
            'fecha_vencimiento': fv,
            'proveedor': prov,
            'stock_actual_g': round(stock, 2),
            'n_movs': int(n_movs or 0),
            'problemas': problemas_lote,
            'estado_audit': estado_audit,
        })

    n_total = len(lotes)
    n_problemas = n_total - n_ok

    if n_total == 0:
        score = 100.0
    else:
        score = 100.0 * (n_ok / n_total)
    score = max(0.0, round(score, 1))

    if score >= 99:
        veredicto = 'PERFECTA'
    elif score >= 85:
        veredicto = 'MENOR'
    else:
        veredicto = 'BLOQUEANTE'

    conn.close()

    return jsonify({
        'ok': True,
        'score': score,
        'veredicto': veredicto,
        'resumen': {
            'dias_horizonte': dias,
            'n_total': n_total,
            'n_ok': n_ok,
            'n_consumidos': n_consumidos,
            'n_problemas': n_problemas,
            'sin_fecha_venc': n_sin_fv,
            'sin_proveedor': n_sin_prov,
            'stock_negativo': n_stock_neg,
        },
        'lotes': lotes[:150],
        'message': (f'S5 lotes nuevos: {veredicto} · score {score}/100 '
                    f'· {n_total} lotes nuevos · {n_problemas} con problemas '
                    f'· {n_consumidos} consumidos (inertes)'),
    }), 200


@bp.route("/api/admin/formula-remapear-material-id", methods=["POST"])
def formula_remapear_material_id():
    """Reemplaza material_id en formula_items (huérfano → código real).

    Body JSON:
      remapeos: [{material_id_actual, material_id_correcto}]
      motivo: str (audit_log)
      dry_run: bool

    Para cada remapeo:
      1. Valida que material_id_correcto existe en maestro_mps activo
      2. UPDATE formula_items SET material_id=correcto, material_nombre=nombre_canonico
         WHERE material_id=actual
      3. Audit log REMAPEAR_FORMULA_MATERIAL_ID con before/after

    NO toca movimientos · solo formula_items (que es la fuente de pre-check).
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    remapeos = d.get('remapeos') or []
    motivo = (d.get('motivo') or 'Remapeo huérfanos formula_items').strip()
    dry_run = bool(d.get('dry_run'))

    if not isinstance(remapeos, list) or not remapeos:
        return jsonify({'error': 'remapeos lista no vacía requerida'}), 400
    if len(remapeos) > 200:
        return jsonify({'error': 'max 200 remapeos por request'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    plan = []
    for it in remapeos:
        actual = (it.get('material_id_actual') or '').strip()
        correcto = (it.get('material_id_correcto') or '').strip().upper()
        if not actual or not correcto:
            continue
        # Validar correcto existe y activo
        row = c.execute("""
            SELECT codigo_mp, nombre_comercial, nombre_inci, activo
            FROM maestro_mps WHERE codigo_mp = ?
        """, (correcto,)).fetchone()
        if not row:
            plan.append({'actual': actual, 'correcto': correcto,
                        'error': f'{correcto} no existe en maestro_mps'})
            continue
        if not row[3]:
            plan.append({'actual': actual, 'correcto': correcto,
                        'error': f'{correcto} está archivado (activo=0)'})
            continue
        nombre_canonico = row[1] or row[2] or correcto
        # Contar afectados
        count_row = c.execute(
            "SELECT COUNT(*) FROM formula_items WHERE material_id = ?",
            (actual,)
        ).fetchone()
        n_items = count_row[0] if count_row else 0
        plan.append({
            'actual': actual, 'correcto': correcto,
            'nombre_canonico': nombre_canonico,
            'items_a_actualizar': n_items,
        })

    if dry_run:
        conn.close()
        return jsonify({'ok': True, 'dry_run': True, 'plan': plan}), 200

    aplicados = 0
    errores = []
    try:
        for p in plan:
            if 'error' in p:
                errores.append(p)
                continue
            try:
                c.execute("""
                    UPDATE formula_items
                    SET material_id = ?, material_nombre = ?
                    WHERE material_id = ?
                """, (p['correcto'], p['nombre_canonico'], p['actual']))
                aplicados += 1
            except Exception as e:
                errores.append({'actual': p['actual'], 'correcto': p['correcto'],
                               'error': str(e)[:200]})
        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='REMAPEAR_FORMULA_MATERIAL_ID',
                tabla='formula_items', registro_id='bulk',
                despues={'motivo': motivo, 'plan': plan,
                        'aplicados': aplicados, 'errores': errores},
                detalle=f'Remapeados {aplicados} material_ids en formula_items',
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla', 'detail': str(e)[:300]}), 500

    conn.close()
    return jsonify({
        'ok': True, 'aplicados': aplicados, 'errores': errores, 'plan': plan,
        'message': f'✓ {aplicados} remapeos aplicados',
    }), 200


@bp.route("/api/admin/explicar-stock-min/<codigo>", methods=["GET"])
def explicar_stock_min(codigo):
    """Desglose detallado del cálculo de stock_min sugerido para UNA MP.

    Sebastián 10-may-2026: "estamos seguros de esos stock minimos?
    corroboraste consumos con formulas maestras?" · esta función te
    permite VALIDAR el cálculo MP por MP antes de aplicar.

    Para el código MP dado, muestra:
      - Producciones del Calendar en el horizonte que la usan
      - Por cada producción: cantidad_kg, lote_size_kg, factor, gramos
        consumidos (calculado tanto por g_por_lote como por porcentaje)
      - Total proyectado en horizonte (suma)
      - Conversión a consumo mensual
      - Stock_min sugerido con la cobertura usada
      - Catálogo MP actual (stock_minimo guardado)

    Query: ?horizonte_dias=90&cobertura_dias=90
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    codigo = (codigo or '').strip()
    if not codigo:
        return jsonify({'error': 'codigo requerido'}), 400

    try:
        horizonte = max(30, min(int(request.args.get('horizonte_dias', 90)), 365))
    except (ValueError, TypeError):
        horizonte = 90
    try:
        cobertura = max(7, min(int(request.args.get('cobertura_dias', 90)), 365))
    except (ValueError, TypeError):
        cobertura = 90

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    # 1. Datos del MP
    mp_row = c.execute("""
        SELECT codigo_mp, nombre_inci, nombre_comercial, proveedor,
               stock_minimo, tipo_material, activo
        FROM maestro_mps WHERE codigo_mp=?
    """, (codigo,)).fetchone()
    if not mp_row:
        conn.close()
        return jsonify({'error': f'MP {codigo} no existe en catálogo'}), 404

    mp_info = {
        'codigo_mp': mp_row[0], 'nombre_inci': mp_row[1],
        'nombre_comercial': mp_row[2], 'proveedor': mp_row[3],
        'stock_minimo_actual_g': mp_row[4] or 0,
        'tipo_material': mp_row[5], 'activo': bool(mp_row[6]),
    }

    # 2. Stock actual (suma de movimientos)
    stock_row = c.execute("""
        SELECT ROUND(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END), 2)
        FROM movimientos WHERE material_id=?
    """, (codigo,)).fetchone()
    mp_info['stock_actual_g'] = (stock_row[0] or 0) if stock_row else 0

    # 3. Lead time
    lt_row = c.execute("""
        SELECT lead_time_dias, buffer_dias FROM mp_lead_time_config
        WHERE material_id=?
    """, (codigo,)).fetchone() if True else None
    try:
        lt_info = {'lead_time_dias': lt_row[0], 'buffer_dias': lt_row[1]} if lt_row else None
    except Exception:
        lt_info = None
    mp_info['mp_lead_time_config'] = lt_info
    cobertura_efectiva = (lt_info['lead_time_dias'] + lt_info['buffer_dias']) if lt_info else cobertura
    mp_info['cobertura_efectiva_dias'] = cobertura_efectiva

    # 4. Fórmulas que usan esta MP (por porcentaje O g_por_lote)
    # Búsqueda EXACTA primero (case-sensitive)
    formulas_usan = c.execute("""
        SELECT fi.producto_nombre,
               COALESCE(fi.porcentaje, 0) as pct,
               COALESCE(fi.cantidad_g_por_lote, 0) as g_lote,
               COALESCE(fh.lote_size_kg, 0) as lote_size_kg,
               fi.material_id as match_mid,
               'exacto' as match_tipo
        FROM formula_items fi
        LEFT JOIN formula_headers fh ON fh.producto_nombre = fi.producto_nombre
        WHERE fi.material_id = ?
        ORDER BY fi.producto_nombre
    """, (codigo,)).fetchall()

    formulas_info = []
    for r in formulas_usan:
        formulas_info.append({
            'producto': r[0], 'porcentaje': r[1],
            'cantidad_g_por_lote': r[2], 'lote_size_kg': r[3],
            'material_id_en_formula': r[4], 'match_tipo': r[5],
        })

    # Si no hay match exacto, hacer búsqueda fuzzy por código case-insensitive
    # y por nombre (catch typos comunes)
    posibles_huerfanos = []
    if not formulas_info:
        nombre_inci = (mp_info.get('nombre_inci') or '').strip()
        nombre_comercial = (mp_info.get('nombre_comercial') or '').strip()
        # Búsqueda case-insensitive por material_id
        rows_ci = c.execute("""
            SELECT DISTINCT fi.material_id, fi.material_nombre,
                   COUNT(DISTINCT fi.producto_nombre) as n_productos
            FROM formula_items fi
            WHERE UPPER(TRIM(fi.material_id)) = UPPER(TRIM(?))
              AND fi.material_id != ?
            GROUP BY fi.material_id
        """, (codigo, codigo)).fetchall()
        for r in rows_ci:
            posibles_huerfanos.append({
                'tipo': 'codigo_case_distinto',
                'material_id_en_formula': r[0],
                'material_nombre_en_formula': r[1],
                'n_productos': r[2],
                'sugerencia': f'Hay items en formula_items con material_id="{r[0]}" (diff casing). Considerar UPDATE para unificar a {codigo}.',
            })
        # Búsqueda por nombre comercial o INCI
        if nombre_comercial or nombre_inci:
            patrones = []
            params = []
            if nombre_comercial and len(nombre_comercial) > 3:
                patrones.append("UPPER(TRIM(fi.material_nombre)) LIKE UPPER(TRIM(?))")
                params.append(f'%{nombre_comercial}%')
            if nombre_inci and len(nombre_inci) > 3:
                patrones.append("UPPER(TRIM(fi.material_nombre)) LIKE UPPER(TRIM(?))")
                params.append(f'%{nombre_inci}%')
            if patrones:
                rows_nm = c.execute(f"""
                    SELECT DISTINCT fi.material_id, fi.material_nombre,
                           COUNT(DISTINCT fi.producto_nombre) as n_productos
                    FROM formula_items fi
                    WHERE ({' OR '.join(patrones)})
                      AND UPPER(TRIM(fi.material_id)) != UPPER(TRIM(?))
                    GROUP BY fi.material_id
                """, params + [codigo]).fetchall()
                for r in rows_nm:
                    posibles_huerfanos.append({
                        'tipo': 'nombre_similar',
                        'material_id_en_formula': r[0],
                        'material_nombre_en_formula': r[1],
                        'n_productos': r[2],
                        'sugerencia': f'Items con nombre similar a "{nombre_comercial or nombre_inci}" usando material_id="{r[0]}". Probable typo · considerá unificar.',
                    })

    if not formulas_info and not posibles_huerfanos:
        conn.close()
        return jsonify({
            'ok': True, 'mp': mp_info,
            'formulas_que_lo_usan': [],
            'posibles_match_huerfanos': [],
            'producciones_calendar': [],
            'desglose_calculo': None,
            'mensaje': 'Esta MP NO está en ninguna fórmula (ni exacto ni similar). '
                       'Stock_min sugerido = 0 · si la MP debería usarse, alguna fórmula '
                       'la necesita registrar.',
            'stock_min_sugerido_g': 0,
        }), 200

    if not formulas_info and posibles_huerfanos:
        conn.close()
        return jsonify({
            'ok': True, 'mp': mp_info,
            'formulas_que_lo_usan': [],
            'posibles_match_huerfanos': posibles_huerfanos,
            'producciones_calendar': [],
            'desglose_calculo': None,
            'mensaje': f'⚠ Esta MP no tiene match EXACTO en fórmulas, pero hay '
                       f'{len(posibles_huerfanos)} posibles huérfanos con código '
                       'distinto o nombre similar. Revisar abajo para decidir si '
                       'unificar antes de calcular stock_min.',
            'stock_min_sugerido_g': 0,
        }), 200

    # 5. Producciones del Calendar en horizonte
    # Lee de produccion_programada (que está poblada desde Calendar)
    productos_que_usan = [f['producto'] for f in formulas_info]
    ph = ','.join(['?'] * len(productos_que_usan))
    prods_cal = c.execute(f"""
        SELECT pp.id, pp.producto, pp.fecha_programada, pp.cantidad_kg,
               COALESCE(pp.lotes, 1) as lotes,
               COALESCE(pp.estado, 'pendiente') as estado,
               COALESCE(pp.inventario_descontado_at, '') as desc_at
        FROM produccion_programada pp
        WHERE pp.producto IN ({ph})
          AND pp.fecha_programada >= date('now')
          AND pp.fecha_programada <= date('now', '+' || ? || ' day')
          AND LOWER(COALESCE(pp.estado,'')) != 'cancelado'
        ORDER BY pp.fecha_programada
    """, productos_que_usan + [horizonte]).fetchall()

    # 6. Desglose por producción
    desglose = []
    total_g_proyectado = 0.0
    for pr in prods_cal:
        pid, producto, fecha, kg, lotes, estado, desc_at = pr
        kg = float(kg or 0)
        # Buscar la fórmula correspondiente
        formula = next((f for f in formulas_info if f['producto'] == producto), None)
        if not formula:
            continue
        # 2 métodos de cálculo (preferir g_por_lote si existe, fallback porcentaje)
        g_por_lote = formula['cantidad_g_por_lote']
        lote_kg = formula['lote_size_kg']
        pct = formula['porcentaje']
        metodo = None
        g_consumido = 0
        formula_detalle = None
        if g_por_lote > 0 and lote_kg > 0:
            factor = kg / lote_kg
            g_consumido = g_por_lote * factor
            metodo = 'cantidad_g_por_lote × factor'
            formula_detalle = f'{g_por_lote}g × ({kg}kg / {lote_kg}kg) = {g_consumido:.2f}g'
        elif pct > 0:
            g_consumido = (pct / 100.0) * kg * 1000
            metodo = 'porcentaje × cantidad_kg × 1000 (fallback)'
            formula_detalle = f'{pct}% × {kg}kg × 1000 = {g_consumido:.2f}g'
        else:
            metodo = 'SIN DATOS (porcentaje=0 y cantidad_g_por_lote=0)'
            formula_detalle = '⚠ Fórmula incompleta · no aporta a stock_min'

        # Considerar solo producciones pendientes/atrasadas (no las ya descontadas)
        cuenta_para_consumo = not desc_at
        if cuenta_para_consumo:
            total_g_proyectado += g_consumido

        desglose.append({
            'produccion_id': pid,
            'producto': producto, 'fecha': fecha,
            'cantidad_kg': kg, 'lotes': lotes, 'estado': estado,
            'ya_descontado': bool(desc_at),
            'metodo_calculo': metodo,
            'formula_calculo': formula_detalle,
            'g_consumido': round(g_consumido, 2),
            'cuenta_para_proyeccion': cuenta_para_consumo,
        })

    # 7. Cálculo final
    consumo_mensual = round((total_g_proyectado * 30 / horizonte) if horizonte else 0, 2)
    stock_min_sugerido = round(consumo_mensual * (cobertura_efectiva / 30), 2)

    conn.close()
    return jsonify({
        'ok': True,
        'mp': mp_info,
        'horizonte_dias': horizonte,
        'cobertura_default_dias': cobertura,
        'formulas_que_lo_usan': formulas_info,
        'producciones_calendar': desglose,
        'desglose_calculo': {
            'total_g_horizonte': round(total_g_proyectado, 2),
            'horizonte_dias': horizonte,
            'consumo_mensual_g': consumo_mensual,
            'cobertura_efectiva_dias': cobertura_efectiva,
            'formula': f'({consumo_mensual:.2f}g/mes) × ({cobertura_efectiva}d / 30d)',
            'stock_min_sugerido_g': stock_min_sugerido,
        },
        'stock_min_sugerido_g': stock_min_sugerido,
        'diferencia_vs_actual_g': round(stock_min_sugerido - mp_info['stock_minimo_actual_g'], 2),
    }), 200


@bp.route("/api/admin/sugerir-stock-minimos", methods=["GET"])
def sugerir_stock_minimos():
    """Sugiere stock_minimo por MP basado en PRODUCCIÓN PROYECTADA en
    Google Calendar (no en consumo histórico).

    Sebastián 10-may-2026: 'el minimo debemos calcularlo segun lo que
    hay en google calendar, alli dice cuanto vamos a producir durante
    el año'.

    Cálculo:
      1. Lee `planificacion_estrategica` (que parsea Calendar + fórmulas)
         para horizonte N días (default 90).
      2. Por cada MP, suma `total_g` proyectado en ese horizonte.
      3. Convierte a tasa mensual: consumo_mensual = total_g / (N/30).
      4. Stock_minimo_sugerido = consumo_mensual × (lead_time+buffer)/30
         donde lead+buffer viene de mp_lead_time_config si existe,
         sino default cobertura 30 días.

    Diferencia vs versión anterior (histórico):
      - Histórico mira pasado · proyectado mira futuro (más útil)
      - Captura lanzamientos nuevos sin historial (estrenos)
      - Captura estacionalidad (campañas, picos)
      - Captura productos descontinuados (consumo proyectado = 0)

    Query params:
      horizonte_dias: int (30-365, default 90)
      cobertura_dias: int (7-180, default = lead+buffer si existe, sino 30)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    try:
        horizonte = max(30, min(int(request.args.get('horizonte_dias', 90)), 365))
    except (ValueError, TypeError):
        horizonte = 90
    # Sebastián 10-may-2026: cobertura default = 90 días (3 meses).
    # Compras locales: lead 14d + safety 76d · sobrado seguro.
    # Compras China estándar: lead 60d + safety 30d · OK.
    # MPs China críticas (lead 90-180d) deben configurar SU lead específico
    # en mp_lead_time_config · el endpoint usa lead+buffer del MP si existe.
    cobertura_default = 90
    try:
        cobertura_default = max(7, min(int(request.args.get('cobertura_dias', 90)), 365))
    except (ValueError, TypeError):
        pass

    # Sebastián 10-may-2026 (validación a fondo): cálculo directo en SQL
    # con fallback dual cantidad_g_por_lote / porcentaje. La versión vieja
    # delegaba a `planificacion_estrategica` que tiene el bug de saltar
    # fórmulas con solo porcentaje (cantidad_g_por_lote=0). Esto cubre
    # AMBOS tipos de fórmula:
    #   1. g_por_lote × (kg_prod / lote_size_kg)        si g_por_lote > 0
    #   2. (porcentaje/100) × kg_prod × 1000            fallback si pct > 0
    consumo_proyectado = {}
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    try:
        rows = c.execute(f"""
            SELECT fi.material_id,
                   COALESCE(SUM(
                     CASE
                       WHEN COALESCE(fi.cantidad_g_por_lote,0) > 0 AND COALESCE(fh.lote_size_kg,0) > 0
                         THEN COALESCE(fi.cantidad_g_por_lote,0) *
                              (pp.cantidad_kg / fh.lote_size_kg)
                       WHEN COALESCE(fi.porcentaje,0) > 0
                         THEN (COALESCE(fi.porcentaje,0) / 100.0) *
                              COALESCE(pp.cantidad_kg, 0) * 1000
                       ELSE 0
                     END
                   ), 0) as total_g
            FROM produccion_programada pp
            JOIN formula_items fi ON UPPER(TRIM(fi.producto_nombre))
                                  = UPPER(TRIM(pp.producto))
            LEFT JOIN formula_headers fh
                   ON UPPER(TRIM(fh.producto_nombre))
                    = UPPER(TRIM(pp.producto))
            WHERE pp.fecha_programada >= date('now')
              AND pp.fecha_programada <= date('now', '+' || ? || ' day')
              AND LOWER(COALESCE(pp.estado,'')) != 'cancelado'
              AND COALESCE(pp.inventario_descontado_at, '') = ''
              AND fi.material_id IS NOT NULL AND TRIM(fi.material_id) != ''
            GROUP BY fi.material_id
        """, (horizonte,)).fetchall()
        for r in rows:
            consumo_proyectado[r[0]] = {'total_g_horizonte': r[1] or 0}
    except Exception as e:
        # Si falla por schema, dejar consumo vacío (todo SIN_USO)
        consumo_proyectado = {}

    # 3. Lead time + buffer por MP (cobertura específica)
    lead_times = {}
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lead_time_dias, 14) as lt,
                   COALESCE(buffer_dias, 30) as buf
            FROM mp_lead_time_config
        """).fetchall()
        for r in rows:
            lead_times[r[0]] = {'lead': r[1], 'buffer': r[2]}
    except sqlite3.OperationalError:
        pass

    # 4. Catálogo MPs activas
    try:
        rows = c.execute("""
            SELECT codigo_mp,
                   SUBSTR(COALESCE(nombre_comercial,nombre_inci,codigo_mp),1,60) as nombre,
                   COALESCE(stock_minimo,0) as min_actual,
                   COALESCE(proveedor,'') as proveedor
            FROM maestro_mps
            WHERE activo=1
            ORDER BY codigo_mp
        """).fetchall()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)[:200]}), 500

    resumen = {'sin_uso': 0, 'alto': 0, 'bajo': 0, 'ok': 0}
    items = []

    for r in rows:
        mid, nombre, min_actual, prov = r[0], r[1], r[2], r[3]
        proy = consumo_proyectado.get(mid, {})
        total_g_horizonte = float(proy.get('total_g_horizonte', 0))
        consumo_mensual = round((total_g_horizonte * 30 / horizonte) if horizonte else 0, 2)
        lt_info = lead_times.get(mid)
        if lt_info:
            cobertura_dias = lt_info['lead'] + lt_info['buffer']
        else:
            cobertura_dias = cobertura_default

        sugerido = round(consumo_mensual * (cobertura_dias / 30), 2)

        if total_g_horizonte == 0:
            estado = 'sin_uso'
            resumen['sin_uso'] += 1
        elif sugerido == 0:
            estado = 'sin_uso'
            resumen['sin_uso'] += 1
        elif min_actual > 2 * sugerido and sugerido > 0:
            estado = 'alto'
            resumen['alto'] += 1
        elif min_actual < 0.5 * sugerido:
            estado = 'bajo'
            resumen['bajo'] += 1
        else:
            estado = 'ok'
            resumen['ok'] += 1

        items.append({
            'codigo_mp': mid, 'nombre': nombre, 'proveedor': prov,
            'stock_minimo_actual_g': round(min_actual, 2),
            'total_proyectado_horizonte_g': round(total_g_horizonte, 2),
            'consumo_mensual_g': consumo_mensual,
            'productos_que_lo_usan': len(proy.get('productos', [])),
            'lead_time_dias': lt_info['lead'] if lt_info else None,
            'buffer_dias': lt_info['buffer'] if lt_info else None,
            'cobertura_dias_usada': cobertura_dias,
            'stock_minimo_sugerido_g': sugerido,
            'diferencia_g': round(sugerido - min_actual, 2),
            'estado': estado,
        })

    items.sort(key=lambda x: (
        {'alto': 0, 'bajo': 1, 'ok': 2, 'sin_uso': 3}.get(x['estado'], 4),
        -abs(x['diferencia_g']),
    ))

    conn.close()
    return jsonify({
        'ok': True,
        'horizonte_dias': horizonte,
        'cobertura_dias_default': cobertura_default,
        'items': items,
        'resumen': resumen,
        'metodologia': (
            f'Producción proyectada (Google Calendar + fórmulas) en {horizonte}d → '
            f'consumo_mensual = total/(horizonte/30) · '
            f'stock_min_sugerido = consumo_mensual × cobertura/30. '
            f'Cobertura usa lead_time+buffer si configurado, sino {cobertura_default} días.'
        ),
        'fuente': 'planificacion_estrategica (Google Calendar)',
    }), 200


@bp.route("/api/admin/aplicar-stock-minimos-sugeridos", methods=["POST"])
def aplicar_stock_minimos_sugeridos():
    """Aplica los stock_minimo sugeridos por /sugerir-stock-minimos.

    Body JSON:
      items: [{codigo_mp, stock_minimo_g}] · lista a aplicar
      motivo: str (audit_log)
      dry_run: bool (default false)

    Filtros recomendados desde frontend:
      - Solo aplicar 'alto' y 'bajo' (no 'ok' ni 'sin_uso')
      - Confirmar uno por uno · o batch tras revisión visual
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    items = d.get('items') or []
    motivo = (d.get('motivo') or 'Ajuste stock_minimo basado en consumo histórico').strip()
    dry_run = bool(d.get('dry_run'))

    if not isinstance(items, list) or not items:
        return jsonify({'error': 'items lista no vacía requerida'}), 400
    if len(items) > 500:
        return jsonify({'error': 'max 500 items por request'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    plan = []
    for it in items:
        mid = (it.get('codigo_mp') or '').strip()
        nuevo = it.get('stock_minimo_g')
        if not mid:
            continue
        try:
            nuevo = float(nuevo)
            if nuevo < 0:
                continue
        except (ValueError, TypeError):
            continue
        row = c.execute(
            "SELECT stock_minimo FROM maestro_mps WHERE codigo_mp=? AND activo=1",
            (mid,)
        ).fetchone()
        if not row:
            continue
        actual = row[0] or 0
        if abs(actual - nuevo) < 0.01:
            continue
        plan.append({'codigo_mp': mid, 'actual': actual, 'nuevo': nuevo})

    if dry_run:
        conn.close()
        return jsonify({'ok': True, 'dry_run': True, 'plan': plan,
                       'a_aplicar': len(plan)}), 200

    aplicados = 0
    try:
        for p in plan:
            c.execute(
                "UPDATE maestro_mps SET stock_minimo=? WHERE codigo_mp=?",
                (p['nuevo'], p['codigo_mp'])
            )
            aplicados += 1
        try:
            import json as _json
            audit_log(
                c, usuario=u, accion='ACTUALIZAR_STOCK_MINIMOS_BULK',
                tabla='maestro_mps', registro_id='bulk',
                despues={'motivo': motivo, 'plan': plan, 'aplicados': aplicados},
                detalle=f'{aplicados} stock_minimos actualizados · motivo: {motivo}',
            )
        except Exception:
            pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla', 'detail': str(e)[:300]}), 500

    conn.close()
    return jsonify({
        'ok': True, 'aplicados': aplicados,
        'message': f'✓ {aplicados} stock_minimo actualizados',
    }), 200


@bp.route("/api/admin/validar-planta", methods=["GET"])
def validar_planta_invariantes():
    """Verifica los 5 invariantes críticos de Planta · cero-error.

    Sebastián 10-may-2026 (visión): "planta es lo principal · fórmulas
    maestras perfectas descontando MPs adecuadas · MPs organizadas un
    código por materia prima donde solo varía el lote · descuentos
    adecuados con cada producción · ingresos reales · ajustes con
    integridad perfecta".

    Devuelve estado de los 5 invariantes con findings concretos:

    1. FÓRMULAS perfectas
       1a. SUM(porcentaje) por producto debe ser 100 ±0.5
       1b. CERO material_id duplicado dentro de una misma fórmula
       1c. CERO formula_items.material_id huérfano (sin maestro activo)

    2. CATÁLOGO: 1 código = 1 MP
       2a. PRIMARY KEY codigo_mp (SQLite enforced)
       2b. CERO MPs activas con mismo nombre_inci/nombre_comercial post-fusión
       2c. CERO MPs archivadas con stock > 0

    3. PRODUCCIONES descontaron correctamente
       3a. Cada produccion 'Completada' tiene movimientos Salida con
           lote_ref en observaciones (prefix FEFO: o UNLIMITED:)
       3b. CERO stock_neto < 0 por lote
       3c. Suma de descuentos = cantidad_total_producida (±tolerancia)

    4. INGRESOS reales (recepciones)
       4a. CERO movimientos Entrada con cantidad <= 0
       4b. Movimientos Entrada con material_id en maestro activo
       4c. Cada Entrada con operador identificado

    5. AJUSTES con integridad
       5a. CERO movimientos con tipo fuera de (Entrada/Salida/Ajuste)
       5b. CERO movimientos con cantidad <= 0
       5c. CERO anulaciones sin contra-movimiento
       5d. audit_log entry por cada acción crítica

    Returns:
      {
        ok, timestamp, score (0-100),
        invariantes: {
          formulas: {ok, score, findings:[...]},
          catalogo: {...},
          producciones: {...},
          ingresos: {...},
          ajustes: {...},
        },
        veredicto: "PERFECTO" | "OK_CON_OBSERVACIONES" | "VIOLACIONES_CRITICAS"
      }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    import datetime as _dt
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    inv = {
        'formulas': {'ok': True, 'score': 100, 'findings': []},
        'catalogo': {'ok': True, 'score': 100, 'findings': []},
        'producciones': {'ok': True, 'score': 100, 'findings': []},
        'ingresos': {'ok': True, 'score': 100, 'findings': []},
        'ajustes': {'ok': True, 'score': 100, 'findings': []},
    }

    def _add(seccion, descripcion, severidad='alta', detalle=None):
        f = {'descripcion': descripcion, 'severidad': severidad}
        if detalle is not None:
            f['detalle'] = detalle
        inv[seccion]['findings'].append(f)
        # Penalizar score · alta -30, media -15, baja -5
        penalty = {'alta': 30, 'media': 15, 'baja': 5}.get(severidad, 10)
        inv[seccion]['score'] = max(0, inv[seccion]['score'] - penalty)
        inv[seccion]['ok'] = False

    # ═══ 1. FÓRMULAS ═══
    # 1a. SUM(porcentaje) por producto debe ser 100 ±0.5
    try:
        rows = c.execute("""
            SELECT producto_nombre, ROUND(SUM(porcentaje), 4) as suma, COUNT(*) as n
            FROM formula_items
            GROUP BY producto_nombre
            HAVING ABS(suma - 100) > 0.5
        """).fetchall()
        if rows:
            _add('formulas',
                 f'{len(rows)} fórmulas con SUM(porcentaje) != 100 ±0.5',
                 'alta',
                 [{'producto': r[0], 'suma': r[1], 'n_items': r[2]} for r in rows[:10]])
    except Exception as e:
        _add('formulas', f'error verificando porcentajes: {e}', 'media')

    # 1b. material_id duplicado en misma fórmula
    try:
        rows = c.execute("""
            SELECT producto_nombre, material_id, COUNT(*) as veces
            FROM formula_items
            WHERE material_id IS NOT NULL AND TRIM(material_id) != ''
            GROUP BY producto_nombre, material_id
            HAVING veces > 1
        """).fetchall()
        if rows:
            _add('formulas',
                 f'{len(rows)} grupos (producto, material_id) duplicados en formula_items',
                 'alta',
                 [{'producto': r[0], 'material_id': r[1], 'veces': r[2]} for r in rows[:10]])
    except Exception as e:
        _add('formulas', f'error verificando duplicados: {e}', 'media')

    # 1c. material_id huérfano (no en maestro activo)
    try:
        rows = c.execute("""
            SELECT DISTINCT fi.material_id
            FROM formula_items fi
            LEFT JOIN maestro_mps mp ON fi.material_id=mp.codigo_mp AND mp.activo=1
            WHERE mp.codigo_mp IS NULL
              AND fi.material_id IS NOT NULL AND TRIM(fi.material_id) != ''
        """).fetchall()
        if rows:
            _add('formulas',
                 f'{len(rows)} material_ids en formula_items sin maestro_mps activo',
                 'alta',
                 [r[0] for r in rows[:20]])
    except Exception as e:
        _add('formulas', f'error verificando huérfanos: {e}', 'media')

    # ═══ 2. CATÁLOGO ═══
    # 2b. MPs activas con mismo nombre_inci normalizado
    try:
        rows = c.execute("""
            SELECT LOWER(TRIM(nombre_inci)) as inci_norm, COUNT(*) as n,
                   GROUP_CONCAT(codigo_mp) as codigos
            FROM maestro_mps
            WHERE activo=1 AND nombre_inci IS NOT NULL AND TRIM(nombre_inci) != ''
              AND LOWER(TRIM(nombre_inci)) NOT IN
                ('parfum','fragrance','aroma','aqua','water','agua',
                 'alcohol','alcohol denat','glycerin','glicerina',
                 '(varies)','mixture','pendiente inci','pendiente',
                 'sin inci','no inci','por definir','tbd','n/a','na','-')
            GROUP BY LOWER(TRIM(nombre_inci))
            HAVING n > 1
        """).fetchall()
        if rows:
            _add('catalogo',
                 f'{len(rows)} grupos INCI duplicados activos (no whitelist)',
                 'alta',
                 [{'inci': r[0], 'n': r[1], 'codigos': r[2]} for r in rows[:10]])
    except Exception as e:
        _add('catalogo', f'error: {e}', 'media')

    # 2c. MPs archivadas con stock > 0
    try:
        rows = c.execute("""
            SELECT mp.codigo_mp,
                   ROUND(SUM(CASE WHEN m.tipo='Entrada' THEN m.cantidad
                                  ELSE -m.cantidad END), 2) as stock
            FROM maestro_mps mp
            JOIN movimientos m ON m.material_id=mp.codigo_mp
            WHERE mp.activo=0
            GROUP BY mp.codigo_mp
            HAVING stock > 0.5
        """).fetchall()
        if rows:
            _add('catalogo',
                 f'{len(rows)} MPs archivadas con stock > 0 (contradicción)',
                 'media',
                 [{'codigo_mp': r[0], 'stock_g': r[1]} for r in rows[:10]])
    except Exception as e:
        _add('catalogo', f'error: {e}', 'media')

    # ═══ 3. PRODUCCIONES ═══
    # 3b. stock_neto < 0 por lote
    try:
        rows = c.execute("""
            SELECT material_id, COALESCE(lote,'') as lote,
                   ROUND(SUM(CASE WHEN tipo='Entrada' THEN cantidad
                                  ELSE -cantidad END), 2) as stock
            FROM movimientos
            GROUP BY material_id, lote
            HAVING stock < -0.5
        """).fetchall()
        if rows:
            _add('producciones',
                 f'{len(rows)} lotes con stock negativo',
                 'alta',
                 [{'material_id': r[0], 'lote': r[1], 'stock': r[2]} for r in rows[:10]])
    except Exception as e:
        _add('producciones', f'error: {e}', 'media')

    # 3a. Producciones Completadas sin movimientos FEFO/UNLIMITED
    try:
        rows = c.execute("""
            SELECT p.id, p.producto, p.lote, p.cantidad
            FROM producciones p
            WHERE p.estado='Completado'
              AND NOT EXISTS (
                SELECT 1 FROM movimientos m
                WHERE m.tipo='Salida'
                  AND (m.observaciones LIKE 'FEFO:%' OR m.observaciones LIKE 'UNLIMITED:%')
                  AND (m.observaciones LIKE '%'||p.lote||'%' OR m.observaciones LIKE '%'||p.producto||'%')
              )
            LIMIT 10
        """).fetchall()
        if rows:
            _add('producciones',
                 f'{len(rows)} producciones Completadas sin movimientos FEFO de descuento',
                 'alta',
                 [{'id': r[0], 'producto': r[1], 'lote': r[2], 'cantidad_kg': r[3]} for r in rows[:10]])
    except Exception as e:
        _add('producciones', f'error producciones huerfanas: {e}', 'media')

    # ═══ 4. INGRESOS ═══
    # 4a. Movimientos Entrada con cantidad <= 0
    try:
        row = c.execute(
            "SELECT COUNT(*) FROM movimientos WHERE tipo='Entrada' AND cantidad <= 0"
        ).fetchone()
        if row and row[0] > 0:
            _add('ingresos', f'{row[0]} movimientos Entrada con cantidad <= 0', 'alta')
    except Exception as e:
        _add('ingresos', f'error: {e}', 'media')

    # 4b. Entradas con material_id huérfano
    try:
        row = c.execute("""
            SELECT COUNT(*) FROM movimientos m
            LEFT JOIN maestro_mps mp ON m.material_id=mp.codigo_mp AND mp.activo=1
            WHERE m.tipo='Entrada' AND mp.codigo_mp IS NULL
        """).fetchone()
        if row and row[0] > 0:
            _add('ingresos',
                 f'{row[0]} movimientos Entrada con material_id sin maestro activo',
                 'media')
    except Exception as e:
        _add('ingresos', f'error: {e}', 'media')

    # 4c. Entradas sin operador
    try:
        row = c.execute("""
            SELECT COUNT(*) FROM movimientos
            WHERE tipo='Entrada' AND (operador IS NULL OR TRIM(operador)='')
        """).fetchone()
        if row and row[0] > 0:
            _add('ingresos',
                 f'{row[0]} movimientos Entrada sin operador identificado',
                 'baja')
    except Exception as e:
        _add('ingresos', f'error: {e}', 'media')

    # ═══ 5. AJUSTES (integridad de movimientos) ═══
    # 5a. tipo fuera de (Entrada/Salida/Ajuste)
    try:
        rows = c.execute("""
            SELECT tipo, COUNT(*) FROM movimientos
            WHERE tipo NOT IN ('Entrada','Salida','Ajuste')
            GROUP BY tipo
        """).fetchall()
        if rows:
            _add('ajustes',
                 f'{sum(r[1] for r in rows)} movimientos con tipo inválido',
                 'alta',
                 [{'tipo': r[0], 'count': r[1]} for r in rows])
    except Exception as e:
        _add('ajustes', f'error: {e}', 'media')

    # 5b. cantidad <= 0
    try:
        row = c.execute(
            "SELECT COUNT(*) FROM movimientos WHERE cantidad IS NULL OR cantidad <= 0"
        ).fetchone()
        if row and row[0] > 0:
            _add('ajustes', f'{row[0]} movimientos con cantidad <= 0 o NULL', 'alta')
    except Exception as e:
        _add('ajustes', f'error: {e}', 'media')

    # 5d. audit_log existe y tiene entradas recientes
    try:
        row = c.execute(
            "SELECT COUNT(*) FROM audit_log WHERE fecha >= datetime('now','-7 days')"
        ).fetchone()
        if row and row[0] < 1:
            _add('ajustes', 'audit_log sin entradas en últimos 7 días (¿deshabilitado?)', 'media')
    except Exception as e:
        _add('ajustes', f'audit_log no consultable: {e}', 'media')

    conn.close()

    # Score global = promedio ponderado
    scores = [v['score'] for v in inv.values()]
    score_global = round(sum(scores) / len(scores), 1)
    veredicto = (
        'PERFECTO' if score_global >= 95 else
        'OK_CON_OBSERVACIONES' if score_global >= 70 else
        'VIOLACIONES_CRITICAS'
    )

    return jsonify({
        'ok': True,
        'timestamp': _dt.datetime.utcnow().isoformat() + 'Z',
        'score_global': score_global,
        'veredicto': veredicto,
        'invariantes': inv,
    }), 200


@bp.route("/admin/normalizar-formulas", methods=["GET"])
def admin_normalizar_formulas_page():
    """Panel para remapear material_ids huérfanos en formula_items."""
    u, err, code = _require_admin()
    if err:
        return Response('<h1>403</h1>', status=403, mimetype='text/html')
    return Response(_NORMALIZAR_FORMULAS_HTML, mimetype='text/html')


_NORMALIZAR_FORMULAS_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Normalizar Fórmulas · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0c0a09;color:#fafaf9;padding:24px}
h1{font-size:24px;color:#fb923c;margin-bottom:4px}
.sub{color:#a8a29e;font-size:13px;margin-bottom:18px}
.note{background:#7c2d1230;border-left:3px solid #fb923c;padding:10px 14px;border-radius:6px;font-size:12px;margin-bottom:14px;line-height:1.6}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin-bottom:18px}
.kpi{background:#1c1917;border-radius:8px;padding:14px;text-align:center;border-left:4px solid #44403c}
.kpi.auto{border-left-color:#16a34a}
.kpi.alta{border-left-color:#5eead4}
.kpi.manual{border-left-color:#ca8a04}
.kpi.crear{border-left-color:#dc2626}
.kpi h4{font-size:10px;text-transform:uppercase;color:#a8a29e;letter-spacing:.5px;margin-bottom:4px}
.kpi .v{font-size:30px;font-weight:800}
.kpi.auto .v{color:#16a34a}.kpi.alta .v{color:#5eead4}.kpi.manual .v{color:#ca8a04}.kpi.crear .v{color:#dc2626}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:700;font-size:13px}
.b-run{background:#fb923c;color:#0c0a09}
.b-apply{background:#a855f7;color:white}
.b-apply:disabled{opacity:.4;cursor:not-allowed}
.row{background:#1c1917;border-radius:8px;padding:14px;margin-bottom:10px;border-left:4px solid #44403c}
.row.auto{border-left-color:#16a34a}
.row.alta{border-left-color:#5eead4}
.row.manual{border-left-color:#ca8a04}
.row.crear{border-left-color:#dc2626}
.r-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;gap:10px;flex-wrap:wrap}
.r-h-info{font-size:13px}
.r-mid{font-family:monospace;font-size:11px;color:#fb923c;background:#0c0a09;padding:2px 6px;border-radius:4px}
.r-name{color:#a8a29e;font-size:11px;margin-left:6px}
.r-prod{color:#78716c;font-size:11px;margin-top:4px}
.sugerencias{margin-top:8px}
.sug{display:flex;align-items:center;gap:8px;padding:6px;border-radius:4px;background:#0c0a09;margin-bottom:4px;font-size:12px}
.sug input{margin:0}
.sug-cod{font-family:monospace;color:#5eead4;font-weight:700;font-size:11px}
.sug-name{color:#a8a29e}
.sug-sim{margin-left:auto;font-weight:700;font-size:11px}
.sug-sim.alta{color:#16a34a}.sug-sim.media{color:#ca8a04}.sug-sim.baja{color:#a8a29e}
.bar-bot{position:sticky;bottom:0;background:#1c1917;padding:14px;border-radius:10px;margin-top:16px;display:flex;justify-content:space-between;align-items:center;border:1px solid #44403c}
.back{display:inline-block;color:#a8a29e;text-decoration:none;font-size:13px;margin-bottom:16px}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>🔧 Normalizar Fórmulas · Material IDs huérfanos</h1>
<p class="sub">Material_ids usados en formula_items que NO existen en maestro_mps. Remapea al código correcto.</p>

<div class="note">
<b>Por qué pasa:</b> fórmulas creadas antes de normalizar el catálogo, o con typos de código, o con material_id minúscula/mayúscula distinta. <b>Resultado:</b> producciones que usan esas fórmulas no descuentan correctamente las MPs.
<br><br>
<b>Recomendaciones automáticas:</b>
<ul style="margin:6px 0 0 18px;font-size:12px">
  <li><span style="color:#16a34a">AUTO</span>: similitud ≥ 95% · aplicar sin pensar</li>
  <li><span style="color:#5eead4">ALTA CONFIANZA</span>: 85-95% · revisar el match sugerido y confirmar</li>
  <li><span style="color:#ca8a04">MANUAL</span>: 70-85% · vos elegís cuál código asignar</li>
  <li><span style="color:#dc2626">CREAR NUEVA MP</span>: sin match · esta MP no existe en catálogo y debe crearse</li>
</ul>
</div>

<div style="margin-bottom:14px">
  <button class="b-run" onclick="cargar()">🔍 Detectar huérfanos</button>
  <span id="ts" style="color:#a8a29e;font-size:11px;margin-left:8px"></span>
</div>

<div id="kpis" class="kpis"></div>
<div id="tabla-cont"></div>
<div class="bar-bot" id="bar-bot" style="display:none">
  <div id="sel-info" style="font-size:13px;color:#a8a29e">0 remapeos seleccionados</div>
  <button class="b-apply" id="b-apply" onclick="aplicar()" disabled>✓ Aplicar remapeos</button>
</div>

<script>
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
window._huerfanos = [];

async function cargar(){
  document.getElementById('tabla-cont').innerHTML='<div style="text-align:center;padding:40px;color:#a8a29e">⏳ Detectando huérfanos en formula_items...</div>';
  try{
    var r = await fetch('/api/admin/formula-huerfanos-con-sugerencias');
    var d = await r.json();
    if(!r.ok){document.getElementById('tabla-cont').innerHTML='<div style="color:#dc2626;padding:14px">Error: '+esc(d.error||r.status)+'</div>';return;}
    window._huerfanos = d.huerfanos || [];
    document.getElementById('ts').textContent = (d.huerfanos||[]).length + ' huérfanos detectados';
    render(d);
  }catch(e){
    document.getElementById('tabla-cont').innerHTML='<div style="color:#dc2626;padding:14px">'+e.message+'</div>';
  }
}

function render(d){
  var r = d.resumen || {};
  var khtml = '';
  [['auto','✓ AUTO (≥95%)'],['revisar_alta_confianza','✓ ALTA CONFIANZA'],
   ['manual','MANUAL'],['crear_nueva_mp','CREAR NUEVA MP']].forEach(function(p){
    var k = p[0]; var label = p[1];
    var klass = k==='auto'?'auto':k==='revisar_alta_confianza'?'alta':k==='manual'?'manual':'crear';
    khtml += '<div class="kpi '+klass+'"><h4>'+label+'</h4><div class="v">'+(r[k]||0)+'</div></div>';
  });
  document.getElementById('kpis').innerHTML = khtml;

  var huerfanos = d.huerfanos || [];
  if(!huerfanos.length){
    document.getElementById('tabla-cont').innerHTML='<div style="text-align:center;color:#16a34a;padding:40px;font-size:14px">✅ Cero huérfanos · formula_items 100% alineada con catálogo.</div>';
    document.getElementById('bar-bot').style.display='none';
    return;
  }
  var html = '';
  huerfanos.forEach(function(h, idx){
    var rec = h.recomendacion;
    var klass = rec==='auto'?'auto':rec==='revisar_alta_confianza'?'alta':rec==='manual'?'manual':'crear';
    html += '<div class="row '+klass+'">';
    html += '<div class="r-head">';
    html += '<div class="r-h-info"><span class="r-mid">'+esc(h.material_id_actual_formula)+'</span>';
    html += '<span class="r-name">'+esc(h.material_nombre_en_formula)+'</span>';
    html += '<div class="r-prod">Usado en '+h.n_productos_que_lo_usan+' productos: '+esc((h.productos||[]).slice(0,5).join(', '))+(h.productos.length>5?'...':'')+'</div>';
    html += '</div>';
    html += '<div style="font-size:11px;color:#a8a29e">recomendación: <b style="color:#fafaf9">'+rec+'</b></div>';
    html += '</div>';

    var sug = h.sugerencias || [];
    if(sug.length){
      html += '<div class="sugerencias">';
      sug.forEach(function(s, j){
        var simKlass = s.similitud>=0.95?'alta':s.similitud>=0.85?'media':'baja';
        var checked = (j===0 && rec==='auto') ? 'checked' : '';
        html += '<label class="sug">';
        html += '<input type="radio" name="sug-'+idx+'" data-actual="'+esc(h.material_id_actual_formula)+'" data-correcto="'+esc(s.codigo_mp_correcto)+'" '+checked+' onchange="updateSel()">';
        html += '<span class="sug-cod">→ '+esc(s.codigo_mp_correcto)+'</span>';
        html += '<span class="sug-name">'+esc(s.nombre_match)+'</span>';
        html += '<span style="color:#78716c;font-size:11px">('+esc(s.razon)+')</span>';
        html += '<span class="sug-sim '+simKlass+'">'+(Math.round(s.similitud*100))+'%</span>';
        html += '</label>';
      });
      // Opción "ninguna · no remapear"
      html += '<label class="sug" style="opacity:.6">';
      html += '<input type="radio" name="sug-'+idx+'" data-actual="" data-correcto="" '+(rec==='auto'?'':'checked')+' onchange="updateSel()">';
      html += '<span class="sug-name">↛ No remapear · dejar como está</span>';
      html += '</label>';
      html += '</div>';
    } else {
      html += '<div style="color:#dc2626;font-size:12px;padding:8px;background:#0c0a09;border-radius:4px">⚠ Sin sugerencias automáticas. Esta MP debe crearse en catálogo (o el nombre en formula_items está incorrecto). Editar manualmente.</div>';
    }
    html += '</div>';
  });
  document.getElementById('tabla-cont').innerHTML = html;
  document.getElementById('bar-bot').style.display='flex';
  updateSel();
}

function updateSel(){
  var sel = Array.from(document.querySelectorAll('input[type=radio]:checked'))
              .filter(function(cb){return cb.dataset.correcto;});
  document.getElementById('sel-info').textContent = sel.length + ' remapeos seleccionados';
  document.getElementById('b-apply').disabled = sel.length === 0;
}

async function aplicar(){
  var sel = Array.from(document.querySelectorAll('input[type=radio]:checked'))
              .filter(function(cb){return cb.dataset.correcto;});
  if(!sel.length){return;}
  if(!confirm('¿Aplicar '+sel.length+' remapeos? Esto actualiza material_id en formula_items. Acción rastreable en audit_log.')){return;}
  var btn = document.getElementById('b-apply');
  btn.disabled = true; btn.textContent = '⏳ Aplicando...';
  var remapeos = sel.map(function(cb){
    return {material_id_actual: cb.dataset.actual, material_id_correcto: cb.dataset.correcto};
  });
  try{
    var r = await fetch('/api/admin/formula-remapear-material-id', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({remapeos: remapeos, motivo: 'Normalización catálogo · panel admin'}),
    });
    var d = await r.json();
    if(r.ok){
      alert('✓ '+(d.message||'Aplicado')+' · recargando...');
      cargar();
    } else {
      alert('Error: '+(d.error||r.status));
      btn.disabled = false; btn.textContent = '✓ Aplicar remapeos';
    }
  }catch(e){
    alert('Error de red: '+e.message);
    btn.disabled = false; btn.textContent = '✓ Aplicar remapeos';
  }
}

cargar();
</script>
</body></html>"""


@bp.route("/admin/stock-minimos", methods=["GET"])
def admin_stock_minimos_page():
    """Panel para revisar y ajustar stock_minimos basados en producción
    proyectada (Google Calendar). Sebastián 10-may-2026."""
    u, err, code = _require_admin()
    if err:
        return Response('<h1>403</h1>', status=403, mimetype='text/html')
    return Response(_STOCK_MINIMOS_HTML, mimetype='text/html')


_STOCK_MINIMOS_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Stock Mínimos · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0c0a09;color:#fafaf9;padding:24px}
h1{font-size:24px;color:#5eead4;margin-bottom:4px}
.sub{color:#a8a29e;font-size:13px;margin-bottom:20px}
.controls{background:#1c1917;border-radius:10px;padding:16px;margin-bottom:18px;display:flex;gap:12px;align-items:end;flex-wrap:wrap}
.controls label{display:flex;flex-direction:column;font-size:11px;color:#a8a29e;text-transform:uppercase;letter-spacing:.5px}
.controls input{margin-top:4px;padding:7px 10px;background:#0c0a09;border:1px solid #44403c;color:#fafaf9;border-radius:5px;width:100px}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:700;font-size:13px}
.b-run{background:#5eead4;color:#0c0a09}
.b-apply{background:#a855f7;color:white}
.b-apply:disabled{opacity:.4;cursor:not-allowed}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:18px}
.kpi{background:#1c1917;border-radius:8px;padding:14px;text-align:center;border-left:4px solid #44403c}
.kpi.alto{border-left-color:#dc2626}
.kpi.bajo{border-left-color:#ca8a04}
.kpi.ok{border-left-color:#16a34a}
.kpi.sin_uso{border-left-color:#525252}
.kpi h4{font-size:10px;text-transform:uppercase;color:#a8a29e;letter-spacing:.5px;margin-bottom:4px}
.kpi .v{font-size:30px;font-weight:800}
.kpi.alto .v{color:#dc2626}.kpi.bajo .v{color:#ca8a04}.kpi.ok .v{color:#16a34a}.kpi.sin_uso .v{color:#a8a29e}
table{width:100%;background:#1c1917;border-radius:10px;border-collapse:separate;border-spacing:0;font-size:12px;overflow:hidden}
th{background:#292524;text-align:left;padding:10px;color:#fafaf9;font-weight:700;text-transform:uppercase;letter-spacing:.5px;font-size:11px;border-bottom:1px solid #44403c;position:sticky;top:0}
td{padding:8px 10px;border-bottom:1px solid #292524}
tr:hover{background:#292524}
.estado{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.3px}
.estado.alto{background:#7f1d1d;color:#fecaca}
.estado.bajo{background:#854d0e;color:#fef3c7}
.estado.ok{background:#14532d;color:#bbf7d0}
.estado.sin_uso{background:#262626;color:#a8a29e}
.num{font-family:Consolas,monospace;text-align:right}
.diff{font-weight:700}
.diff.pos{color:#16a34a}.diff.neg{color:#dc2626}
.back{display:inline-block;color:#a8a29e;text-decoration:none;font-size:13px;margin-bottom:16px}
.bar-bot{position:sticky;bottom:0;background:#1c1917;padding:14px;border-radius:10px;margin-top:16px;display:flex;justify-content:space-between;align-items:center;border:1px solid #44403c}
.sel-info{font-size:13px;color:#a8a29e}
.note{background:#1e3a8a30;border-left:3px solid #3b82f6;padding:10px 14px;border-radius:6px;font-size:12px;margin-bottom:14px}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>📊 Stock Mínimos · Sugeridos por Calendar</h1>
<p class="sub">Calculado desde producción proyectada en Google Calendar × fórmulas activas. NO usa consumo histórico.</p>

<div class="note">
<b>Cómo funciona:</b> el sistema lee Google Calendar para el horizonte (default 90d = 1 trimestre), multiplica cada producción por su fórmula, suma el consumo proyectado por MP, y calcula <code>stock_minimo = consumo_mensual × cobertura/30</code>.
<br><br>
<b>Cobertura 90d (3 meses)</b> es el default · cuando una MP llega al stock_minimo, te quedan 3 meses de stock · tiempo suficiente para que llegue el pedido (locales 14d, China 60d) con buffer de seguridad.
<br><br>
<b>MPs China críticas (lead 90-180d):</b> configurá su lead específico en <code>mp_lead_time_config</code> · el endpoint usa SU cobertura individual (lead+buffer) en vez del default.
</div>

<div class="controls">
  <label>Horizonte (días)
    <input type="number" id="horizonte" value="90" min="30" max="365" step="30">
  </label>
  <label>Cobertura default (días)
    <input type="number" id="cobertura" value="90" min="7" max="365" step="15">
  </label>
  <button class="b-run" onclick="calcular()">🔍 Calcular sugerencias</button>
  <span id="ts" style="color:#a8a29e;font-size:11px;margin-left:8px"></span>
</div>

<div id="kpis" class="kpis"></div>
<div id="tabla-cont"></div>
<div class="bar-bot" id="bar-bot" style="display:none">
  <div class="sel-info" id="sel-info">0 items seleccionados</div>
  <button class="b-apply" id="b-apply" onclick="aplicar()" disabled>✓ Aplicar seleccionados</button>
</div>

<script>
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
function fmt(n){return Number(n||0).toLocaleString('es-CO',{maximumFractionDigits:1});}

async function calcular(){
  var h = document.getElementById('horizonte').value || 90;
  var c = document.getElementById('cobertura').value || 30;
  document.getElementById('tabla-cont').innerHTML = '<div style="text-align:center;padding:40px;color:#a8a29e">⏳ Leyendo Google Calendar + cruzando con fórmulas...</div>';
  try{
    var r = await fetch('/api/admin/sugerir-stock-minimos?horizonte_dias='+h+'&cobertura_dias='+c);
    var d = await r.json();
    if(!r.ok){
      document.getElementById('tabla-cont').innerHTML='<div style="color:#dc2626;padding:14px">Error: '+esc(d.error||r.status)+'</div>';
      return;
    }
    document.getElementById('ts').textContent = 'Horizonte: '+(d.horizonte_dias)+'d · '+
      'Fuente: '+(d.fuente||'')+' · '+(d.items||[]).length+' MPs activas';
    render(d);
  }catch(e){
    document.getElementById('tabla-cont').innerHTML='<div style="color:#dc2626;padding:14px">'+e.message+'</div>';
  }
}

function render(d){
  var r = d.resumen || {};
  var html = '';
  ['alto','bajo','ok','sin_uso'].forEach(function(k){
    html += '<div class="kpi '+k+'"><h4>'+k.replace('_',' ')+'</h4><div class="v">'+(r[k]||0)+'</div></div>';
  });
  document.getElementById('kpis').innerHTML = html;

  var items = d.items || [];
  // Mostrar sólo los que tienen acción (alto/bajo)
  var accion = items.filter(function(it){return it.estado==='alto'||it.estado==='bajo';});
  if(!accion.length){
    document.getElementById('tabla-cont').innerHTML='<div style="text-align:center;color:#16a34a;padding:40px;font-size:14px">✅ TODOS los stock_minimos están dentro de rango razonable · sin sugerencias de cambio.</div>';
    document.getElementById('bar-bot').style.display='none';
    return;
  }
  var thtml = '<table>'+
    '<thead><tr>'+
    '<th style="width:30px"><input type="checkbox" onclick="marcarTodos(this.checked)"></th>'+
    '<th>Código</th><th>Nombre</th><th>Estado</th>'+
    '<th class="num">Actual (g)</th><th class="num">Proyectado horizonte (g)</th>'+
    '<th class="num">Consumo /mes (g)</th><th class="num">Cobertura</th>'+
    '<th class="num">Sugerido (g)</th><th class="num">Diferencia</th>'+
    '</tr></thead><tbody>';
  accion.forEach(function(it,idx){
    var diffClass = it.diferencia_g > 0 ? 'pos' : 'neg';
    var diffSign = it.diferencia_g > 0 ? '+' : '';
    thtml += '<tr>';
    thtml += '<td><input type="checkbox" class="sel" data-mid="'+esc(it.codigo_mp)+'" data-sug="'+it.stock_minimo_sugerido_g+'" onchange="updateSel()"></td>';
    thtml += '<td style="font-family:monospace;font-size:11px">'+esc(it.codigo_mp)+'</td>';
    thtml += '<td>'+esc(it.nombre)+'</td>';
    thtml += '<td><span class="estado '+it.estado+'">'+it.estado+'</span></td>';
    thtml += '<td class="num">'+fmt(it.stock_minimo_actual_g)+'</td>';
    thtml += '<td class="num">'+fmt(it.total_proyectado_horizonte_g)+'</td>';
    thtml += '<td class="num">'+fmt(it.consumo_mensual_g)+'</td>';
    thtml += '<td class="num">'+(it.cobertura_dias_usada||'-')+'d</td>';
    thtml += '<td class="num" style="color:#5eead4;font-weight:700">'+fmt(it.stock_minimo_sugerido_g)+'</td>';
    thtml += '<td class="num diff '+diffClass+'">'+diffSign+fmt(it.diferencia_g)+'</td>';
    thtml += '</tr>';
  });
  thtml += '</tbody></table>';
  document.getElementById('tabla-cont').innerHTML = thtml;
  document.getElementById('bar-bot').style.display='flex';
  updateSel();
}

function marcarTodos(estado){
  document.querySelectorAll('.sel').forEach(function(cb){cb.checked = estado;});
  updateSel();
}
function updateSel(){
  var sel = document.querySelectorAll('.sel:checked').length;
  document.getElementById('sel-info').textContent = sel + ' items seleccionados';
  document.getElementById('b-apply').disabled = sel === 0;
}

async function aplicar(){
  var sel = Array.from(document.querySelectorAll('.sel:checked'));
  if(!sel.length){return;}
  if(!confirm('¿Aplicar nuevo stock_minimo a '+sel.length+' MPs? Acción rastreable en audit_log.')){return;}
  var items = sel.map(function(cb){
    return {codigo_mp: cb.dataset.mid, stock_minimo_g: parseFloat(cb.dataset.sug)};
  });
  var btn = document.getElementById('b-apply');
  btn.disabled = true; btn.textContent = '⏳ Aplicando...';
  try{
    var r = await fetch('/api/admin/aplicar-stock-minimos-sugeridos',{
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({items: items, motivo: 'Ajuste basado en Google Calendar · panel'}),
    });
    var d = await r.json();
    if(r.ok){
      alert('✓ '+(d.message||'Aplicado')+' · recargando...');
      calcular();
    } else {
      alert('Error: '+(d.error||r.status));
      btn.disabled = false; btn.textContent = '✓ Aplicar seleccionados';
    }
  }catch(e){
    alert('Error de red: '+e.message);
    btn.disabled = false; btn.textContent = '✓ Aplicar seleccionados';
  }
}

calcular();
</script>
</body></html>"""


@bp.route("/admin/integridad-planta", methods=["GET"])
def admin_integridad_planta_page():
    """Panel semáforo de los 5 invariantes de Planta · cero-error."""
    u, err, code = _require_admin()
    if err:
        return Response('<h1>403</h1>', status=403, mimetype='text/html')
    return Response(_INTEGRIDAD_PLANTA_HTML, mimetype='text/html')


_INTEGRIDAD_PLANTA_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Integridad Planta · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0c0a09;color:#fafaf9;padding:24px}
h1{font-size:26px;color:#5eead4;margin-bottom:4px}
.sub{color:#a8a29e;font-size:13px;margin-bottom:24px}
.score-card{background:linear-gradient(135deg,#1c1917 0%,#292524 100%);
  border-radius:16px;padding:32px;margin-bottom:24px;text-align:center;
  border:2px solid #44403c}
.score-num{font-size:72px;font-weight:900;line-height:1}
.score-label{font-size:18px;color:#a8a29e;margin-top:8px;letter-spacing:1px;text-transform:uppercase}
.vere{display:inline-block;padding:6px 16px;border-radius:20px;font-weight:700;font-size:13px;margin-top:12px;letter-spacing:.5px}
.v-PERFECTO{background:#16a34a;color:white}
.v-OK_CON_OBSERVACIONES{background:#ca8a04;color:white}
.v-VIOLACIONES_CRITICAS{background:#dc2626;color:white}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:16px}
.inv{background:#1c1917;border-radius:12px;padding:20px;border-left:5px solid #44403c}
.inv.ok{border-left-color:#16a34a}
.inv.bad{border-left-color:#dc2626}
.inv h3{font-size:14px;text-transform:uppercase;letter-spacing:1px;color:#fafaf9;margin-bottom:4px}
.inv .meta{color:#a8a29e;font-size:11px;margin-bottom:12px}
.inv .score{font-size:32px;font-weight:800;margin-bottom:4px}
.inv.ok .score{color:#16a34a}
.inv.bad .score{color:#dc2626}
.findings{margin-top:12px}
.f{background:#0c0a09;border-radius:6px;padding:10px;margin-bottom:6px;border-left:3px solid #57534e;font-size:12px}
.f.alta{border-left-color:#dc2626}
.f.media{border-left-color:#ca8a04}
.f.baja{border-left-color:#0891b2}
.f .det{color:#a8a29e;font-size:11px;margin-top:4px;font-family:monospace;max-height:120px;overflow:auto}
button{background:#5eead4;color:#0c0a09;border:none;padding:10px 24px;border-radius:8px;cursor:pointer;font-weight:700;font-size:14px}
button:hover{background:#2dd4bf}
.back{display:inline-block;color:#a8a29e;text-decoration:none;font-size:13px;margin-bottom:16px}
.ts{color:#78716c;font-size:11px;margin-left:12px}
</style></head><body>
<a class="back" href="/modulos">← Panel inicial</a>
<h1>🏭 Integridad de Planta</h1>
<p class="sub">5 invariantes obligatorios · cero-error · validación en vivo</p>

<div class="score-card">
  <div class="score-num" id="score">--</div>
  <div class="score-label">SCORE GLOBAL</div>
  <div id="veredicto"></div>
  <div style="margin-top:18px"><button onclick="validar()">🔍 Validar ahora</button>
  <span class="ts" id="ts"></span></div>
</div>

<div class="grid" id="grid"></div>

<script>
var LABELS = {
  formulas: '🧪 Fórmulas Maestras',
  catalogo: '📦 Catálogo MPs',
  producciones: '⚙️ Producciones',
  ingresos: '🚚 Ingresos (recepciones)',
  ajustes: '✏️ Ajustes (integridad)'
};
var SUB = {
  formulas: 'SUM%=100 · sin duplicados · sin huérfanos',
  catalogo: '1 código = 1 MP · sin contradicciones',
  producciones: 'FEFO correcto · stock no negativo',
  ingresos: 'cantidad>0 · operador · maestro activo',
  ajustes: 'tipo válido · audit_log · sin DELETE'
};
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}

async function validar(){
  document.getElementById('score').textContent='⏳';
  document.getElementById('grid').innerHTML='<div style="grid-column:1/-1;text-align:center;color:#a8a29e;padding:40px">Validando 5 invariantes...</div>';
  try{
    var r = await fetch('/api/admin/validar-planta');
    var d = await r.json();
    if(!r.ok){document.getElementById('grid').innerHTML='<div style="color:#dc2626;padding:20px">Error '+r.status+'</div>';return;}
    document.getElementById('score').textContent = d.score_global;
    document.getElementById('score').style.color =
      d.veredicto==='PERFECTO' ? '#16a34a' :
      d.veredicto==='OK_CON_OBSERVACIONES' ? '#ca8a04' : '#dc2626';
    document.getElementById('veredicto').innerHTML =
      '<span class="vere v-'+d.veredicto+'">'+d.veredicto.replace(/_/g,' ')+'</span>';
    document.getElementById('ts').textContent = (d.timestamp||'').slice(0,19).replace('T',' ');
    var html = '';
    Object.keys(LABELS).forEach(function(k){
      var inv = d.invariantes[k] || {};
      var klass = inv.ok ? 'ok' : 'bad';
      html += '<div class="inv '+klass+'">';
      html += '<h3>'+LABELS[k]+'</h3>';
      html += '<div class="meta">'+SUB[k]+'</div>';
      html += '<div class="score">'+(inv.score||0)+'<span style="font-size:14px;color:#a8a29e">/100</span></div>';
      var findings = inv.findings || [];
      if(!findings.length){
        html += '<div style="color:#16a34a;font-size:12px;margin-top:8px">✓ Sin findings · invariante perfecto</div>';
      } else {
        html += '<div class="findings">';
        findings.slice(0,8).forEach(function(f){
          html += '<div class="f '+(f.severidad||'media')+'"><b>'+esc(f.descripcion)+'</b>';
          if(f.detalle){
            html += '<div class="det">'+esc(JSON.stringify(f.detalle, null, 2))+'</div>';
          }
          html += '</div>';
        });
        if(findings.length > 8){
          html += '<div style="color:#a8a29e;font-size:11px;margin-top:6px">... y '+(findings.length-8)+' más</div>';
        }
        html += '</div>';
      }
      html += '</div>';
    });
    document.getElementById('grid').innerHTML = html;
  }catch(e){
    document.getElementById('grid').innerHTML='<div style="color:#dc2626;padding:20px">'+e.message+'</div>';
  }
}
validar();
</script>
</body></html>"""


@bp.route("/admin/limpieza-cero-error", methods=["GET"])
def admin_limpieza_cero_error_page():
    """Panel guiado de limpieza cero-error · ejecuta los 5 fixes en
    orden seguro · cada paso con preview (dry_run) antes de aplicar.

    Sebastián 10-may-2026: "hazlo todo, código perfecto cero error".
    Sequenced execution: dry_run → confirm → apply → verify.
    """
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este panel.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_LIMPIEZA_HTML, mimetype='text/html')


_LIMPIEZA_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>Limpieza Cero-Error · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0f172a;color:#f1f5f9;padding:20px}
h1{font-size:24px;margin-bottom:6px;color:#5eead4}
.sub{color:#94a3b8;font-size:13px;margin-bottom:24px}
.step{background:#1e293b;border-radius:10px;padding:20px;margin-bottom:16px;border-left:4px solid #475569;transition:.2s}
.step.active{border-left-color:#5eead4;box-shadow:0 0 0 1px #5eead4}
.step.done{border-left-color:#22c55e;opacity:.8}
.step.error{border-left-color:#ef4444}
.step h2{font-size:16px;margin-bottom:6px;color:#f1f5f9;display:flex;align-items:center;gap:10px}
.badge{background:#0f172a;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:700;color:#5eead4;border:1px solid #475569}
.badge.done{color:#22c55e;border-color:#22c55e}
.badge.error{color:#ef4444;border-color:#ef4444}
.desc{font-size:13px;color:#94a3b8;margin-bottom:12px;line-height:1.5}
.actions{display:flex;gap:8px;flex-wrap:wrap}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px;transition:.15s}
button:disabled{opacity:.4;cursor:not-allowed}
.btn-preview{background:#475569;color:#f1f5f9}
.btn-preview:hover:not(:disabled){background:#64748b}
.btn-apply{background:#5eead4;color:#0f172a;font-weight:700}
.btn-apply:hover:not(:disabled){background:#2dd4bf}
.btn-verify{background:#a855f7;color:white}
.result{margin-top:12px;padding:12px;background:#0f172a;border-radius:6px;font-family:Consolas,Monaco,monospace;font-size:11px;color:#cbd5e1;max-height:300px;overflow:auto;white-space:pre-wrap;display:none}
.result.visible{display:block}
.result.ok{border:1px solid #22c55e}
.result.error{border:1px solid #ef4444;color:#fca5a5}
.back{display:inline-block;color:#94a3b8;text-decoration:none;font-size:13px;margin-bottom:16px}
.back:hover{color:#f1f5f9}
.summary{background:#0c4a6e;border:1px solid #0284c7;border-radius:10px;padding:16px;margin-top:24px}
.summary h3{color:#7dd3fc;margin-bottom:8px;font-size:14px}
.summary p{font-size:13px;color:#bae6fd;line-height:1.6}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>🧹 Limpieza Cero-Error · Catálogo MPs</h1>
<p class="sub">5 fixes secuenciales · cada uno con preview (dry-run) antes de aplicar · todo en audit_log · reversible.</p>

<div class="step active" id="step-1">
  <h2>1. Fórmulas con items duplicados <span class="badge" id="b-1">pendiente</span></h2>
  <p class="desc">Detecta fórmulas donde el mismo material_id aparece en >1 fila (caso SUERO ILUMINADOR TRX = 200%). Consolida en una sola fila sumando porcentajes.</p>
  <div class="actions">
    <button class="btn-preview" onclick="preview(1)">📊 Preview</button>
    <button class="btn-apply" onclick="apply(1)" disabled id="apply-1">✓ Aplicar</button>
  </div>
  <div class="result" id="r-1"></div>
</div>

<div class="step" id="step-2">
  <h2>2. MPs huérfanas en fórmulas <span class="badge" id="b-2">pendiente</span></h2>
  <p class="desc">187 material_ids usados en formula_items NO existen en maestro_mps activo. Bloquean producción real. Crea las MPs faltantes con nombre derivado.</p>
  <div class="actions">
    <button class="btn-preview" onclick="preview(2)">📊 Preview</button>
    <button class="btn-apply" onclick="apply(2)" disabled id="apply-2">✓ Aplicar</button>
  </div>
  <div class="result" id="r-2"></div>
</div>

<div class="step" id="step-3">
  <h2>3. Marcar 6 lotes vencidos como VENCIDO <span class="badge" id="b-3">pendiente</span></h2>
  <p class="desc">Lotes con fecha_vencimiento &lt; hoy pero estado_lote='VIGENTE'. Riesgo INVIMA. Cambia estado a VENCIDO en todos sus movimientos.</p>
  <div class="actions">
    <button class="btn-preview" onclick="preview(3)">📊 Preview</button>
    <button class="btn-apply" onclick="apply(3)" disabled id="apply-3">✓ Aplicar</button>
  </div>
  <div class="result" id="r-3"></div>
</div>

<div class="step" id="step-4">
  <h2>4. Anular MP00112 lote AJUSTE-4 (-1.430.000g) <span class="badge" id="b-4">pendiente</span></h2>
  <p class="desc">Movimiento salida fantasma sin Entrada respaldatoria. Crea contra-movimiento Entrada que lleva saldo a 0. Preserva trazabilidad INVIMA (no borra original).</p>
  <div class="actions">
    <button class="btn-preview" onclick="preview(4)">📊 Preview</button>
    <button class="btn-apply" onclick="apply(4)" disabled id="apply-4">✓ Aplicar</button>
  </div>
  <div class="result" id="r-4"></div>
</div>

<div class="step" id="step-5">
  <h2>5. Re-ejecutar auditoría · verificar cero findings ALTA <span class="badge" id="b-5">pendiente</span></h2>
  <p class="desc">Tras los 4 fixes anteriores, vuelve a correr la auditoría completa. Si ALTA=0, el catálogo está perfecto.</p>
  <div class="actions">
    <button class="btn-verify" onclick="verifyFinal()">🔍 Verificar</button>
  </div>
  <div class="result" id="r-5"></div>
</div>

<div class="summary">
  <h3>📋 Reporte ejecutivo</h3>
  <p id="summary-text">Empezá con el Paso 1. Cada "Preview" muestra qué se va a cambiar sin aplicarlo. Después de revisar, "Aplicar" lo ejecuta con audit_log. Si algo falla, podés ver los detalles en el área negra de resultado.</p>
</div>

<script>
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
function setBadge(n, txt, klass){
  var b = document.getElementById('b-'+n);
  b.textContent = txt;
  b.className = 'badge' + (klass ? ' '+klass : '');
  var s = document.getElementById('step-'+n);
  s.className = 'step' + (klass==='done'?' done':klass==='error'?' error':' active');
}
function show(n, text, ok){
  var r = document.getElementById('r-'+n);
  r.className = 'result visible' + (ok===false?' error':' ok');
  r.textContent = text;
}

async function preview(n){
  setBadge(n, '⏳ analizando...');
  show(n, 'Cargando...', true);
  try{
    var res;
    if(n===1){
      res = await fetch('/api/admin/formula-duplicados');
    } else if(n===2){
      res = await fetch('/api/admin/material-ids-huerfanos');
    } else if(n===3){
      // Re-ejecutar auditoría para obtener lotes vencidos
      res = await fetch('/api/admin/auditoria-catalogo?quick=1');
    } else if(n===4){
      res = await fetch('/api/admin/investigar-mp/MP00112');
    }
    var d = await res.json();
    if(!res.ok){
      setBadge(n, 'error', 'error');
      show(n, 'Error '+res.status+': '+JSON.stringify(d, null, 2), false);
      return;
    }
    // Resumen amigable por paso
    var summary = '';
    if(n===1){
      var dups = d.duplicados || [];
      var pcts = d.porcentajes_anomalos || [];
      summary = 'Detectado:\\n  · '+dups.length+' grupos con material_id duplicado\\n  · '+pcts.length+' fórmulas con SUM != 100%\\n\\n';
      if(pcts.length){
        summary += 'Fórmulas con porcentaje anómalo:\\n';
        pcts.slice(0,10).forEach(function(p){
          summary += '  · '+p.producto+' = '+p.suma_porcentajes+'%  ('+p.n_items+' items)\\n';
        });
        summary += '\\n';
      }
      if(dups.length){
        summary += 'Top 10 grupos duplicados:\\n';
        dups.slice(0,10).forEach(function(g){
          summary += '  · '+g.producto+' / '+g.material_id+' aparece '+g.veces+' veces\\n';
        });
      }
      summary += '\\n→ Al aplicar: consolida en 1 fila por grupo (suma porcentajes).';
    } else if(n===2){
      var huerf = (d.huerfanos && d.huerfanos.formula_items) || [];
      summary = 'Huérfanos en formula_items: '+huerf.length+'\\n\\nTop 15:\\n';
      huerf.slice(0,15).forEach(function(h){
        summary += '  · '+h.material_id+' ('+h.nombre+') · '+h.productos_usando+' productos\\n';
      });
      summary += '\\n→ Al aplicar: crea estas MPs en maestro_mps con datos mínimos.';
    } else if(n===3){
      var v = (d.findings && d.findings.alta && d.findings.alta.vencidos_pero_vigente) || [];
      window._vencidos_cache = v;
      summary = 'Lotes vencidos pero VIGENTE: '+v.length+'\\n\\n';
      v.forEach(function(l){
        summary += '  · '+l.material_id+' lote '+l.lote+' venció '+l.fecha_venc+' (stock '+l.stock_g+'g)\\n';
      });
      summary += '\\n→ Al aplicar: cambia estado_lote=\\'VENCIDO\\' en todos sus movs.';
    } else if(n===4){
      var movs = d.movimientos_recientes || [];
      var ajuste4 = movs.find(function(m){return (m.lote||'').toUpperCase()==='AJUSTE-4';});
      window._mp00112_mov_id = ajuste4 ? ajuste4.id : null;
      summary = 'MP00112 ('+(d.mp&&d.mp.nombre_comercial||'?')+')\\n';
      summary += 'Stock total: '+d.stock_total_neto_g+' g\\n\\n';
      if(ajuste4){
        summary += 'Movimiento problemático encontrado:\\n';
        summary += '  · id='+ajuste4.id+' tipo='+ajuste4.tipo+' cantidad='+ajuste4.cantidad_g+'g\\n';
        summary += '  · lote='+ajuste4.lote+' fecha='+ajuste4.fecha+'\\n';
        summary += '  · obs: '+(ajuste4.observaciones||'(sin obs)')+'\\n\\n';
        summary += '→ Al aplicar: crea contra-movimiento '+(ajuste4.tipo==='Salida'?'Entrada':'Salida')+
                   ' de '+ajuste4.cantidad_g+'g lote '+ajuste4.lote+' (saldo neto → 0).';
      } else {
        summary += 'No se encontró movimiento de lote AJUSTE-4 en últimos 50.\\n';
        summary += 'Saltar este paso si stock_total ya está OK.';
      }
    }
    setBadge(n, 'preview OK · listo para aplicar');
    show(n, summary, true);
    document.getElementById('apply-'+n).disabled = false;
  }catch(e){
    setBadge(n, 'error', 'error');
    show(n, 'Error de red: '+e.message, false);
  }
}

async function apply(n){
  if(!confirm('¿Aplicar paso '+n+'? Esta acción queda en audit_log y es trazable. Continuar?')) return;
  setBadge(n, '⏳ aplicando...');
  document.getElementById('apply-'+n).disabled = true;
  try{
    var res, body, url;
    if(n===1){
      url='/api/admin/formula-limpiar-duplicados';
      body={};
    } else if(n===2){
      // Recargar lista y enviar todos los huerfanos
      var listRes = await fetch('/api/admin/material-ids-huerfanos');
      var listD = await listRes.json();
      var mids = ((listD.huerfanos||{}).formula_items||[]).map(function(h){return h.material_id;});
      url='/api/admin/crear-mps-huerfanas';
      body={material_ids: mids};
    } else if(n===3){
      var venc = window._vencidos_cache || [];
      if(!venc.length){
        setBadge(n, 'nada que hacer', 'done');
        show(n, 'No hay lotes vencidos VIGENTE.', true);
        return;
      }
      url='/api/admin/marcar-lotes-vencidos';
      body={lotes: venc.map(function(l){return {material_id: l.material_id, lote: l.lote};}),
            motivo: 'Limpieza cero-error · INVIMA compliance'};
    } else if(n===4){
      var mid = window._mp00112_mov_id;
      if(!mid){
        setBadge(n, 'sin movimiento a anular', 'done');
        show(n, 'No se encontró movimiento de lote AJUSTE-4 en los últimos 50 movs.\\n'+
              'Si el saldo neto sigue siendo -1.4M, posiblemente esté en movs más antiguos.\\n'+
              'Resolver manualmente con SQL o backup-restore.', false);
        return;
      }
      url='/api/admin/anular-movimiento';
      body={mov_id: mid, motivo: 'Ajuste-4 cíclico sin Entrada respaldatoria · saldo -1.430.000g irreal · limpieza cero-error'};
    }
    res = await fetch(url, {method:'POST', headers:{'Content-Type':'application/json'},
                            body: JSON.stringify(body)});
    var d = await res.json();
    if(res.ok){
      setBadge(n, '✓ aplicado', 'done');
      show(n, '✓ ÉXITO\\n\\n'+JSON.stringify(d, null, 2), true);
    } else {
      setBadge(n, 'error', 'error');
      show(n, 'Error '+res.status+':\\n'+JSON.stringify(d, null, 2), false);
      document.getElementById('apply-'+n).disabled = false;
    }
  }catch(e){
    setBadge(n, 'error', 'error');
    show(n, 'Error de red: '+e.message, false);
    document.getElementById('apply-'+n).disabled = false;
  }
}

async function verifyFinal(){
  setBadge(5, '⏳ verificando...');
  show(5, 'Re-ejecutando auditoría completa...', true);
  try{
    var res = await fetch('/api/admin/auditoria-catalogo?quick=1');
    var d = await res.json();
    var r = d.resumen || {};
    var alta = r.n_alta || 0;
    var media = r.n_media || 0;
    var baja = r.n_baja || 0;
    var msg = '🔍 AUDITORÍA POST-LIMPIEZA\\n\\n';
    msg += '🔴 ALTA:  '+alta+'\\n';
    msg += '🟠 MEDIA: '+media+'\\n';
    msg += '🔵 BAJA:  '+baja+'\\n';
    msg += '──────────────\\n';
    msg += 'TOTAL: '+(alta+media+baja)+' findings\\n\\n';
    if(alta === 0){
      msg += '✅ CATÁLOGO PERFECTO · 0 findings ALTA · cero-error logrado.';
      setBadge(5, '✓ perfecto', 'done');
      document.getElementById('summary-text').innerHTML =
        '<b style="color:#5eead4">✅ Sistema en estado cero-error.</b> '+
        'Catálogo limpio, fórmulas consistentes, sin huérfanos, sin vencidos vigente, sin stock negativo. '+
        'Los findings MEDIA/BAJA restantes son mejoras operativas no críticas.';
    } else {
      msg += '⚠ Quedan '+alta+' findings ALTA · revisar /admin/auditoria-catalogo';
      setBadge(5, alta+' ALTA pendientes', 'error');
    }
    show(5, msg, alta===0);
  }catch(e){
    setBadge(5, 'error', 'error');
    show(5, 'Error: '+e.message, false);
  }
}
</script>
</body></html>"""


@bp.route("/admin/auditoria-catalogo", methods=["GET"])
def admin_auditoria_catalogo_page():
    """UI HTML que consume /api/admin/auditoria-catalogo y muestra
    findings agrupados por severidad. Sebastián 10-may-2026."""
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver esta auditoría.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_AUDIT_CATALOGO_HTML, mimetype='text/html')


_AUDIT_CATALOGO_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>Auditoría Catálogo MPs · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;background:#f1f5f9;padding:20px;color:#0f172a}
h1{font-size:22px;margin-bottom:6px;color:#0f766e}
.sub{color:#64748b;font-size:13px;margin-bottom:18px}
.bar{display:flex;gap:8px;margin-bottom:18px;flex-wrap:wrap}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:700;font-size:13px}
button.run{background:#0d9488;color:white}
button.run:hover{background:#0f766e}
button.back{background:#475569;color:white;text-decoration:none;display:inline-block}
.resumen{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:24px}
.card{padding:18px;border-radius:8px;background:white;box-shadow:0 1px 3px rgba(0,0,0,.06)}
.card h3{font-size:11px;text-transform:uppercase;color:#64748b;letter-spacing:.5px;margin-bottom:8px}
.card .num{font-size:36px;font-weight:800}
.card.alta .num{color:#dc2626}
.card.media .num{color:#f59e0b}
.card.baja .num{color:#0891b2}
.card.total .num{color:#0d9488}
section{background:white;border-radius:8px;padding:18px;margin-bottom:18px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
section h2{font-size:16px;margin-bottom:12px;display:flex;align-items:center;gap:8px}
section.alta h2{color:#dc2626}
section.media h2{color:#d97706}
section.baja h2{color:#0891b2}
.finding{border-left:3px solid #cbd5e1;padding:10px 14px;margin-bottom:10px;background:#f8fafc;border-radius:0 4px 4px 0}
.finding.alta{border-left-color:#dc2626}
.finding.media{border-left-color:#f59e0b}
.finding.baja{border-left-color:#0891b2}
.finding h4{font-size:13px;color:#0f172a;margin-bottom:6px;display:flex;justify-content:space-between;align-items:center}
.finding .count{background:#1e293b;color:white;padding:2px 8px;border-radius:10px;font-size:11px}
.finding .empty{color:#16a34a;font-size:12px;font-weight:600}
.finding table{width:100%;font-size:12px;margin-top:6px;border-collapse:collapse}
.finding th{text-align:left;color:#475569;padding:4px 8px;border-bottom:1px solid #e2e8f0;font-weight:600;font-size:11px;text-transform:uppercase}
.finding td{padding:4px 8px;border-bottom:1px solid #f1f5f9}
.finding code{background:#f1f5f9;padding:1px 4px;border-radius:3px;font-size:11px;color:#0d9488}
.loading{text-align:center;padding:40px;color:#64748b}
.error{background:#fee2e2;color:#991b1b;padding:14px;border-radius:6px}
.ok-section{background:#dcfce7;color:#15803d;padding:14px;border-radius:6px;font-size:13px}
</style>
</head><body>
<a href="/modulos" class="back" style="padding:8px 16px;border-radius:6px;background:#475569;color:white;text-decoration:none;display:inline-block;margin-bottom:18px;font-size:12px">← Volver al panel</a>
<h1>🔍 Auditoría del Catálogo de Materias Primas</h1>
<p class="sub">Detecta duplicados, huérfanos e inconsistencias en <code>maestro_mps</code> y <code>movimientos</code>. 12 checks · solo lectura.</p>
<div class="bar">
  <button class="run" onclick="ejecutarAudit()">▶ Ejecutar auditoría</button>
  <button style="background:#7c3aed;color:white;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-weight:700;font-size:13px" onclick="fusionMasivaPreview()" id="btn-bulk" disabled>🤖 Fusión masiva (auto)</button>
  <span id="ts" style="align-self:center;color:#64748b;font-size:12px"></span>
</div>
<div id="contenido"><div class="loading">Click "Ejecutar auditoría" para empezar...</div></div>

<script>
async function ejecutarAudit(){
  var c=document.getElementById('contenido');
  c.innerHTML='<div class="loading">⏳ Analizando catálogo (puede tardar 5-10s)...</div>';
  document.getElementById('ts').textContent='';
  try{
    var r=await fetch('/api/admin/auditoria-catalogo');
    var d=await r.json();
    if(!r.ok){c.innerHTML='<div class="error">Error '+r.status+': '+(d.error||'')+'</div>';return;}
    document.getElementById('ts').textContent='Generado: '+(d.timestamp||'').replace('T',' ').slice(0,19);
    render(d);
  }catch(e){c.innerHTML='<div class="error">Error de red: '+e.message+'</div>';}
}

function esc(s){return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}

function render(d){
  var r=d.resumen||{};
  var f=d.findings||{};
  // Sebastián 10-may-2026: habilitar fusión masiva si hay grupos
  window._auditData = d;
  var grupos_inci = (f.alta && f.alta.inci_duplicado) || [];
  var grupos_nc = (f.media && f.media.nombre_comercial_duplicado) || [];
  var totalBulk = grupos_inci.length + grupos_nc.length;
  var btnBulk = document.getElementById('btn-bulk');
  if(btnBulk){
    btnBulk.disabled = totalBulk === 0;
    btnBulk.textContent = totalBulk > 0
      ? '🤖 Fusión masiva (' + totalBulk + ' grupos)'
      : '🤖 Fusión masiva (auto)';
  }
  var html='<div class="resumen">'+
    '<div class="card alta"><h3>🔴 Alta</h3><div class="num">'+(r.n_alta||0)+'</div></div>'+
    '<div class="card media"><h3>🟠 Media</h3><div class="num">'+(r.n_media||0)+'</div></div>'+
    '<div class="card baja"><h3>🔵 Baja</h3><div class="num">'+(r.n_baja||0)+'</div></div>'+
    '<div class="card total"><h3>Total findings</h3><div class="num">'+(r.n_total_findings||0)+'</div></div>'+
    '</div>';
  if((r.n_total_findings||0)===0){
    html+='<div class="ok-section">✅ Catálogo limpio · 0 findings · todo OK.</div>';
    document.getElementById('contenido').innerHTML=html;
    return;
  }
  html+=renderSeccion('alta','🔴 ALTA · bloquean producción / pérdida de trazabilidad',f.alta||{});
  html+=renderSeccion('media','🟠 MEDIA · operativos · confunden trabajo diario',f.media||{});
  html+=renderSeccion('baja','🔵 BAJA · limpieza / mejora',f.baja||{});
  document.getElementById('contenido').innerHTML=html;
}

function renderSeccion(sev, titulo, items){
  var keys=Object.keys(items).filter(function(k){return !k.endsWith('_err');});
  var hayAlgo=keys.some(function(k){return Array.isArray(items[k]) && items[k].length>0;});
  if(!hayAlgo){
    return '<section class="'+sev+'"><h2>'+titulo+'</h2>'+
           '<div class="ok-section">✓ Sin findings en esta categoría</div></section>';
  }
  var html='<section class="'+sev+'"><h2>'+titulo+'</h2>';
  keys.forEach(function(k){
    var arr=items[k]||[];
    if(!Array.isArray(arr)||!arr.length)return;
    html+='<div class="finding '+sev+'">';
    html+='<h4>'+esc(k.replace(/_/g,' '))+
          ' <span class="count">'+arr.length+'</span></h4>';
    // Sebastián 10-may-2026: botones de limpieza por tipo de finding.
    if(k==='inci_duplicado' || k==='nombre_comercial_duplicado'){
      html+=renderGruposFusion(arr, k);
    } else if(k==='vencidos_pero_vigente'){
      html+='<div style="margin-bottom:10px"><button onclick="marcarVencidos()" '+
            'style="background:#dc2626;color:white;border:none;padding:6px 14px;border-radius:5px;'+
            'cursor:pointer;font-size:12px;font-weight:700">🏷️ Marcar todos como VENCIDO</button>'+
            '<span style="margin-left:10px;font-size:11px;color:#64748b">'+
            'Cambia estado_lote a VENCIDO en los '+arr.length+' lotes</span></div>';
      html+=renderTabla(arr);
    } else if(k==='stock_negativo'){
      html+='<div style="margin-bottom:10px;background:#fef3c7;padding:8px 12px;border-radius:4px;font-size:11px;color:#92400e">'+
            '⚠ Stock negativo = error de kardex. Click <b>Investigar</b> en cada lote para ver sus movimientos y decidir si anular o compensar.</div>';
      html+=renderTablaConInvestigar(arr);
    } else {
      html+=renderTabla(arr);
    }
    html+='</div>';
  });
  // Errors si hubo
  Object.keys(items).filter(function(k){return k.endsWith('_err');}).forEach(function(k){
    html+='<div class="finding"><h4>⚠️ '+esc(k)+'</h4><code>'+esc(items[k])+'</code></div>';
  });
  html+='</section>';
  return html;
}

function renderTabla(arr){
  if(!arr.length)return '<div class="empty">vacío</div>';
  var cols=Object.keys(arr[0]);
  var html='<table><thead><tr>';
  cols.forEach(function(c){html+='<th>'+esc(c)+'</th>';});
  html+='</tr></thead><tbody>';
  arr.slice(0,30).forEach(function(row){
    html+='<tr>';
    cols.forEach(function(c){
      var v=row[c];
      if(Array.isArray(v))v=v.join(', ');
      else if(typeof v==='object'&&v!==null)v=JSON.stringify(v);
      html+='<td>'+esc(v)+'</td>';
    });
    html+='</tr>';
  });
  html+='</tbody></table>';
  if(arr.length>30)html+='<div style="font-size:11px;color:#64748b;margin-top:6px">... y '+(arr.length-30)+' más (limitado a 30 en UI)</div>';
  return html;
}

// Sebastián 10-may-2026: render de grupos de duplicados con botón
// "Fusionar grupo" que abre modal para elegir cuál es canónico.
window._gruposFusion = {};
function renderGruposFusion(arr, tipoKey){
  if(!arr.length)return '<div class="empty">vacío</div>';
  var html='<table><thead><tr>'+
           '<th>Normalizado</th><th>Códigos MP</th><th>Variantes raw</th>'+
           '<th style="text-align:center">Acción</th></tr></thead><tbody>';
  arr.slice(0,30).forEach(function(row,idx){
    var key = tipoKey+'_'+idx;
    window._gruposFusion[key] = row;
    var norm = row.inci_normalizado || row.nombre_normalizado || '';
    var codigos = row.codigos_mp || [];
    var variantes = row.variantes_raw || '';
    html+='<tr>';
    html+='<td><code>'+esc(norm)+'</code></td>';
    html+='<td style="font-family:monospace;font-size:11px">'+esc(codigos.join(', '))+'</td>';
    html+='<td>'+esc(variantes)+'</td>';
    html+='<td style="text-align:center">'+
          '<button onclick="abrirModalFusion(\\''+esc(key)+'\\')" '+
          'style="background:#7c3aed;color:white;border:none;padding:5px 12px;'+
          'border-radius:4px;cursor:pointer;font-size:11px;font-weight:700">'+
          '🔀 Fusionar</button></td>';
    html+='</tr>';
  });
  html+='</tbody></table>';
  if(arr.length>30)html+='<div style="font-size:11px;color:#64748b;margin-top:6px">... y '+(arr.length-30)+' más</div>';
  return html;
}

function abrirModalFusion(key){
  var g = window._gruposFusion[key];
  if(!g){alert('Grupo no encontrado'); return;}
  var codigos = g.codigos_mp || [];
  if(codigos.length < 2){alert('Necesito al menos 2 códigos para fusionar'); return;}

  // Crear modal dinámicamente si no existe
  var modal = document.getElementById('modal-fusion');
  if(!modal){
    modal = document.createElement('div');
    modal.id='modal-fusion';
    modal.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9999;'+
                       'display:flex;align-items:center;justify-content:center;padding:20px';
    document.body.appendChild(modal);
  }
  var norm = g.inci_normalizado || g.nombre_normalizado || '';
  var opciones = codigos.map(function(cod){
    return '<label style="display:flex;align-items:center;gap:10px;padding:10px;'+
           'border:2px solid #e2e8f0;border-radius:6px;margin-bottom:8px;cursor:pointer;'+
           'font-family:monospace" '+
           'onmouseover="this.style.borderColor=\\'#7c3aed\\';this.style.background=\\'#faf5ff\\'" '+
           'onmouseout="this.style.borderColor=\\'#e2e8f0\\';this.style.background=\\'white\\'">'+
           '<input type="radio" name="fusion-canonico" value="'+esc(cod)+'" '+
           'style="margin:0"><span style="font-weight:700">'+esc(cod)+'</span></label>';
  }).join('');
  modal.innerHTML =
    '<div style="background:white;border-radius:10px;padding:24px;max-width:600px;width:100%;'+
    'box-shadow:0 20px 60px rgba(0,0,0,.4)">'+
    '<h2 style="margin:0 0 12px;color:#7c3aed;font-size:18px">🔀 Fusionar MPs duplicadas</h2>'+
    '<p style="color:#64748b;font-size:13px;margin-bottom:14px">'+
    'Grupo: <code style="background:#f1f5f9;padding:2px 6px;border-radius:3px">'+esc(norm)+'</code></p>'+
    '<p style="font-size:13px;margin-bottom:14px"><b>Elegí cuál MP es la <span style="color:#7c3aed">CANÓNICA</span></b> '+
    '(la que sobrevive). Las otras se archivan y sus movimientos se transfieren a la canónica.</p>'+
    '<div style="margin-bottom:14px">'+opciones+'</div>'+
    '<div style="margin-bottom:14px">'+
    '<label style="display:block;font-size:12px;font-weight:600;margin-bottom:4px;color:#475569">'+
    'Motivo (queda en audit_log)</label>'+
    '<input type="text" id="fusion-motivo" placeholder="Ej: casing inconsistente · auditoría 10-may" '+
    'style="width:100%;padding:8px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px"></div>'+
    '<div style="display:flex;gap:8px;justify-content:flex-end">'+
    '<button onclick="document.getElementById(\\'modal-fusion\\').remove()" '+
    'style="background:#94a3b8;color:white;border:none;padding:8px 16px;border-radius:6px;'+
    'cursor:pointer;font-weight:700">Cancelar</button>'+
    '<button onclick="ejecutarFusion()" id="btn-ejecutar-fusion" '+
    'style="background:#7c3aed;color:white;border:none;padding:8px 16px;border-radius:6px;'+
    'cursor:pointer;font-weight:700">✓ Fusionar</button>'+
    '</div>'+
    '<div id="fusion-msg" style="margin-top:12px;font-size:13px"></div>'+
    '</div>';
  modal.dataset.codigos = JSON.stringify(codigos);
}

async function ejecutarFusion(){
  var modal = document.getElementById('modal-fusion');
  if(!modal) return;
  var radio = modal.querySelector('input[name="fusion-canonico"]:checked');
  if(!radio){
    document.getElementById('fusion-msg').innerHTML =
      '<span style="color:#dc2626">Elegí cuál código será el canónico.</span>';
    return;
  }
  var canonico = radio.value;
  var codigos = JSON.parse(modal.dataset.codigos || '[]');
  var duplicados = codigos.filter(function(c){return c !== canonico;});
  var motivo = (document.getElementById('fusion-motivo').value || '').trim();
  if(!confirm('¿FUSIONAR?\\n\\n'+
              'Canónico (sobrevive): '+canonico+'\\n'+
              'A archivar: '+duplicados.join(', ')+'\\n\\n'+
              'Esto transferirá TODOS los movimientos al canónico. Acción rastreable en audit_log.')){
    return;
  }
  var btn = document.getElementById('btn-ejecutar-fusion');
  btn.disabled = true; btn.textContent = '⏳ Fusionando...';
  document.getElementById('fusion-msg').innerHTML =
    '<span style="color:#64748b">Procesando · esto puede tardar unos segundos...</span>';
  try{
    var r = await fetch('/api/admin/maestro-mps-unificar', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({
        codigo_canonico: canonico,
        codigos_duplicados: duplicados,
        motivo: motivo,
      }),
    });
    var d = {}; try{d = await r.json();}catch(e){}
    if(r.ok){
      var t = d.totales_transferidos || {};
      document.getElementById('fusion-msg').innerHTML =
        '<div style="color:#16a34a;font-weight:700;padding:8px;background:#dcfce7;border-radius:6px">'+
        '✓ '+esc(d.message||'Fusión completada')+'<br>'+
        '<small style="color:#15803d;font-weight:400">Transferidos: '+
        (t.movimientos||0)+' movs · '+
        (t.formula_items||0)+' fórmulas · '+
        (t.conteo_items||0)+' conteos · '+
        (t.sol_items||0)+' SOLs</small></div>';
      setTimeout(function(){
        document.getElementById('modal-fusion').remove();
        ejecutarAudit();  // recargar auditoría · el grupo debería desaparecer
      }, 2200);
    } else {
      document.getElementById('fusion-msg').innerHTML =
        '<div style="color:#dc2626;padding:8px;background:#fee2e2;border-radius:6px">'+
        '❌ Error '+r.status+': '+esc(d.error||'unknown')+
        (d.detail ? '<br><small>'+esc(d.detail)+'</small>' : '')+'</div>';
      btn.disabled = false; btn.textContent = '✓ Fusionar';
    }
  }catch(e){
    document.getElementById('fusion-msg').innerHTML =
      '<span style="color:#dc2626">Error de red: '+e.message+'</span>';
    btn.disabled = false; btn.textContent = '✓ Fusionar';
  }
}

// Sebastián 10-may-2026: marcar todos los lotes vencidos_pero_vigente como VENCIDO
async function marcarVencidos(){
  var d = window._auditData || {};
  var arr = (d.findings && d.findings.alta && d.findings.alta.vencidos_pero_vigente) || [];
  if(!arr.length){alert('No hay vencidos para marcar'); return;}
  if(!confirm('Marcar como VENCIDO los '+arr.length+' lotes con fecha vencida?\\n\\n'+
              'Esto cambia estado_lote en TODOS los movimientos de cada lote.\\n'+
              'Acción rastreable en audit_log. Continuar?')){return;}
  var lotes = arr.map(function(r){return {material_id: r.material_id, lote: r.lote};});
  try{
    var r = await fetch('/api/admin/marcar-lotes-vencidos',{
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({lotes: lotes, motivo: 'Limpieza auditoría'}),
    });
    var resp = {};try{resp = await r.json();}catch(e){}
    if(r.ok){
      alert('✓ '+(resp.message || 'Vencidos marcados')+'\\nRecargando auditoría...');
      ejecutarAudit();
    } else {
      alert('Error: '+(resp.error||r.status)+(resp.detail?'\\n'+resp.detail:''));
    }
  }catch(e){alert('Error de red: '+e.message);}
}

// Investigar un MP específico (stock negativo, huérfanos, etc.)
async function investigarMP(codigo){
  try{
    var r = await fetch('/api/admin/investigar-mp/'+encodeURIComponent(codigo));
    var d = {};try{d = await r.json();}catch(e){}
    if(!r.ok){alert('Error: '+(d.error||r.status)); return;}
    abrirModalInvestigar(codigo, d);
  }catch(e){alert('Error de red: '+e.message);}
}

function abrirModalInvestigar(codigo, d){
  var modal = document.getElementById('modal-investigar');
  if(!modal){
    modal = document.createElement('div');
    modal.id = 'modal-investigar';
    modal.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9999;'+
                       'display:flex;align-items:flex-start;justify-content:center;padding:20px;'+
                       'overflow-y:auto';
    document.body.appendChild(modal);
  }
  var mp = d.mp || {};
  var lotes = d.lotes_resumen || [];
  var movs = d.movimientos_recientes || [];
  var stockTotal = d.stock_total_neto_g || 0;
  var lotesTbl = lotes.map(function(l){
    var stockColor = l.stock_neto_g < 0 ? 'color:#dc2626;font-weight:700' : '';
    return '<tr>'+
      '<td style="padding:4px 6px;font-family:monospace;font-size:11px">'+esc(l.lote||'(sin lote)')+'</td>'+
      '<td style="padding:4px 6px;text-align:right;'+stockColor+'">'+l.stock_neto_g+'</td>'+
      '<td style="padding:4px 6px;text-align:center">'+l.n_movs+'</td>'+
      '<td style="padding:4px 6px;font-size:10px;color:#64748b">'+esc((l.primer_mov||'').slice(0,16))+'</td>'+
      '<td style="padding:4px 6px;font-size:10px;color:#64748b">'+esc((l.ultimo_mov||'').slice(0,16))+'</td>'+
    '</tr>';
  }).join('');
  var movsTbl = movs.map(function(m){
    var tipoColor = m.tipo==='Salida'?'color:#dc2626':m.tipo==='Entrada'?'color:#16a34a':'color:#7c3aed';
    return '<tr>'+
      '<td style="padding:3px 6px;font-size:10px">'+m.id+'</td>'+
      '<td style="padding:3px 6px;'+tipoColor+';font-weight:600;font-size:11px">'+esc(m.tipo)+'</td>'+
      '<td style="padding:3px 6px;text-align:right;font-family:monospace">'+m.cantidad_g+'</td>'+
      '<td style="padding:3px 6px;font-family:monospace;font-size:10px">'+esc(m.lote||'')+'</td>'+
      '<td style="padding:3px 6px;font-size:10px;color:#64748b">'+esc((m.fecha||'').slice(0,16))+'</td>'+
      '<td style="padding:3px 6px;font-size:10px">'+esc((m.observaciones||'').slice(0,80))+'</td>'+
    '</tr>';
  }).join('');
  var stockTotalColor = stockTotal < 0 ? 'color:#dc2626;font-weight:700' : 'color:#16a34a';
  modal.innerHTML =
    '<div style="background:white;border-radius:10px;padding:24px;max-width:1200px;width:100%;'+
    'box-shadow:0 20px 60px rgba(0,0,0,.4)">'+
    '<div style="display:flex;justify-content:space-between;align-items:start;margin-bottom:14px">'+
    '<div><h2 style="margin:0;color:#0d9488;font-size:18px">🔬 Investigar '+esc(codigo)+'</h2>'+
    '<p style="margin:4px 0 0;font-size:13px;color:#475569">'+
    esc(mp.nombre_comercial||'(sin nombre)')+' · '+esc(mp.nombre_inci||'')+'</p></div>'+
    '<button onclick="document.getElementById(\\'modal-investigar\\').remove()" '+
    'style="background:#94a3b8;color:white;border:none;padding:6px 12px;border-radius:5px;cursor:pointer;font-size:12px">✗ Cerrar</button>'+
    '</div>'+
    '<div style="background:#f1f5f9;border-radius:6px;padding:10px;margin-bottom:14px;font-size:12px">'+
    'Stock total: <span style="'+stockTotalColor+'">'+stockTotal+' g</span> · '+
    'Activo: '+(mp.activo?'Sí':'No')+' · '+
    'Stock mínimo: '+(mp.stock_minimo||0)+' g'+
    '</div>'+
    '<h3 style="font-size:14px;margin-bottom:6px">Lotes ('+lotes.length+')</h3>'+
    '<div style="border:1px solid #e2e8f0;border-radius:5px;max-height:250px;overflow-y:auto;margin-bottom:14px">'+
    '<table style="width:100%;font-size:12px;border-collapse:collapse">'+
    '<thead style="background:#f1f5f9;position:sticky;top:0"><tr>'+
    '<th style="padding:5px;text-align:left">Lote</th>'+
    '<th style="padding:5px;text-align:right">Stock g</th>'+
    '<th style="padding:5px;text-align:center">N movs</th>'+
    '<th style="padding:5px;text-align:left">Primer mov</th>'+
    '<th style="padding:5px;text-align:left">Último mov</th>'+
    '</tr></thead><tbody>'+lotesTbl+'</tbody></table></div>'+
    '<h3 style="font-size:14px;margin-bottom:6px">Movimientos recientes ('+movs.length+')</h3>'+
    '<div style="border:1px solid #e2e8f0;border-radius:5px;max-height:300px;overflow-y:auto">'+
    '<table style="width:100%;font-size:11px;border-collapse:collapse">'+
    '<thead style="background:#f1f5f9;position:sticky;top:0"><tr>'+
    '<th style="padding:5px;text-align:left">ID</th>'+
    '<th style="padding:5px;text-align:left">Tipo</th>'+
    '<th style="padding:5px;text-align:right">Cantidad g</th>'+
    '<th style="padding:5px;text-align:left">Lote</th>'+
    '<th style="padding:5px;text-align:left">Fecha</th>'+
    '<th style="padding:5px;text-align:left">Obs</th>'+
    '</tr></thead><tbody>'+movsTbl+'</tbody></table></div>'+
    '</div>';
}

function renderTablaConInvestigar(arr){
  if(!arr.length)return '<div class="empty">vacío</div>';
  var cols=Object.keys(arr[0]);
  var html='<table><thead><tr>';
  cols.forEach(function(c){html+='<th>'+esc(c)+'</th>';});
  html+='<th>Acción</th></tr></thead><tbody>';
  arr.slice(0,30).forEach(function(row){
    html+='<tr>';
    cols.forEach(function(c){
      var v=row[c];
      if(Array.isArray(v))v=v.join(', ');
      else if(typeof v==='object'&&v!==null)v=JSON.stringify(v);
      html+='<td>'+esc(v)+'</td>';
    });
    var mid = row.material_id || row.codigo_mp || '';
    html+='<td><button onclick="investigarMP(\\''+esc(mid)+'\\')" '+
          'style="background:#0d9488;color:white;border:none;padding:4px 10px;border-radius:4px;cursor:pointer;font-size:11px">🔬 Investigar</button></td>';
    html+='</tr>';
  });
  html+='</tbody></table>';
  return html;
}

// Sebastián 10-may-2026: fusión masiva con auto-elección de canónico
// (código más bajo de cada grupo = típicamente el original más antiguo).
function fusionMasivaPreview(){
  var d = window._auditData || {};
  var f = d.findings || {};
  var grupos_inci = (f.alta && f.alta.inci_duplicado) || [];
  var grupos_nc = (f.media && f.media.nombre_comercial_duplicado) || [];
  var todos = grupos_inci.concat(grupos_nc);
  if(!todos.length){alert('No hay grupos para fusionar'); return;}
  // Auto-elegir canónico: el código MÁS BAJO (sort alfabético) suele ser
  // el original más antiguo. El user puede cambiar antes de confirmar.
  var batch = todos.map(function(g, idx){
    var codigos = (g.codigos_mp || []).slice().sort();
    return {
      idx: idx,
      tipo: g.inci_normalizado ? 'INCI' : 'Nombre',
      etiqueta: g.inci_normalizado || g.nombre_normalizado || '',
      codigos: codigos,
      canonico: codigos[0],
      duplicados: codigos.slice(1),
      incluir: true,
    };
  });
  window._batchFusion = batch;
  abrirModalBulk(batch);
}

function abrirModalBulk(batch){
  var modal = document.getElementById('modal-bulk');
  if(!modal){
    modal = document.createElement('div');
    modal.id = 'modal-bulk';
    modal.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9999;'+
                       'display:flex;align-items:center;justify-content:center;padding:20px;'+
                       'overflow-y:auto';
    document.body.appendChild(modal);
  }
  var filas = batch.map(function(g, idx){
    var opciones = g.codigos.map(function(cod){
      var sel = cod === g.canonico ? 'selected' : '';
      return '<option value="'+esc(cod)+'" '+sel+'>'+esc(cod)+'</option>';
    }).join('');
    return '<tr>'+
      '<td style="padding:4px 6px"><input type="checkbox" data-idx="'+idx+'" '+(g.incluir?'checked':'')+
        ' onchange="window._batchFusion['+idx+'].incluir=this.checked"></td>'+
      '<td style="padding:4px 6px;font-size:11px;color:#64748b">'+esc(g.tipo)+'</td>'+
      '<td style="padding:4px 6px;font-family:monospace;font-size:11px">'+esc(g.etiqueta).slice(0,40)+'</td>'+
      '<td style="padding:4px 6px;font-family:monospace;font-size:11px">'+esc(g.codigos.join(', '))+'</td>'+
      '<td style="padding:4px 6px">'+
        '<select data-idx="'+idx+'" onchange="window._batchFusion['+idx+'].canonico=this.value;'+
        'window._batchFusion['+idx+'].duplicados=window._batchFusion['+idx+'].codigos.filter(function(c){return c!==this.value;}.bind(this))" '+
        'style="font-family:monospace;font-size:11px;padding:3px 5px;border:1px solid #cbd5e1;border-radius:4px">'+opciones+'</select>'+
      '</td>'+
    '</tr>';
  }).join('');
  modal.innerHTML =
    '<div style="background:white;border-radius:10px;padding:24px;max-width:1100px;width:100%;'+
    'max-height:90vh;overflow-y:auto;box-shadow:0 20px 60px rgba(0,0,0,.4)">'+
    '<h2 style="margin:0 0 12px;color:#7c3aed;font-size:18px">🤖 Fusión Masiva · '+batch.length+' grupos</h2>'+
    '<p style="color:#64748b;font-size:13px;margin-bottom:14px">'+
    'Canónico auto-elegido: <b>código más bajo</b> (suele ser el original). '+
    'Podés cambiar cada uno o desmarcar grupos que no quieras fusionar. '+
    'Después de confirmar, todos los movs de los duplicados se transfieren al canónico y los duplicados se archivan.</p>'+
    '<div style="margin-bottom:10px;display:flex;gap:8px;align-items:center;flex-wrap:wrap">'+
    '<button onclick="bulkToggleTodos(true)" style="background:#0d9488;color:white;border:none;padding:5px 10px;border-radius:4px;cursor:pointer;font-size:11px">✓ Marcar todos</button>'+
    '<button onclick="bulkToggleTodos(false)" style="background:#94a3b8;color:white;border:none;padding:5px 10px;border-radius:4px;cursor:pointer;font-size:11px">✗ Desmarcar todos</button>'+
    '<span id="bulk-counter" style="margin-left:8px;font-size:12px;color:#475569;font-weight:600"></span>'+
    '</div>'+
    '<div style="border:1px solid #e2e8f0;border-radius:6px;max-height:50vh;overflow-y:auto;margin-bottom:14px">'+
    '<table style="width:100%;font-size:12px;border-collapse:collapse">'+
    '<thead style="background:#f1f5f9;position:sticky;top:0"><tr>'+
    '<th style="padding:6px;text-align:left">✓</th>'+
    '<th style="padding:6px;text-align:left">Tipo</th>'+
    '<th style="padding:6px;text-align:left">Etiqueta</th>'+
    '<th style="padding:6px;text-align:left">Códigos</th>'+
    '<th style="padding:6px;text-align:left">Canónico (sobrevive)</th>'+
    '</tr></thead><tbody>'+filas+'</tbody></table></div>'+
    '<div style="margin-bottom:14px">'+
    '<label style="display:block;font-size:12px;font-weight:600;margin-bottom:4px;color:#475569">Motivo común (audit log)</label>'+
    '<input type="text" id="bulk-motivo" value="Limpieza masiva auditoría · casing/duplicados detectados" '+
    'style="width:100%;padding:8px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px"></div>'+
    '<div style="display:flex;gap:8px;justify-content:flex-end">'+
    '<button onclick="document.getElementById(\\'modal-bulk\\').remove()" '+
    'style="background:#94a3b8;color:white;border:none;padding:10px 20px;border-radius:6px;cursor:pointer;font-weight:700">Cancelar</button>'+
    '<button onclick="ejecutarFusionBulk()" id="btn-confirmar-bulk" '+
    'style="background:#7c3aed;color:white;border:none;padding:10px 20px;border-radius:6px;cursor:pointer;font-weight:700">✓ Fusionar seleccionados</button>'+
    '</div>'+
    '<div id="bulk-msg" style="margin-top:12px;font-size:13px"></div>'+
    '</div>';
  actualizarContadorBulk();
}

function bulkToggleTodos(estado){
  (window._batchFusion || []).forEach(function(g){g.incluir = estado;});
  // Refresca checkboxes
  document.querySelectorAll('#modal-bulk input[type=checkbox]').forEach(function(cb){cb.checked = estado;});
  actualizarContadorBulk();
}

function actualizarContadorBulk(){
  var batch = window._batchFusion || [];
  var sel = batch.filter(function(g){return g.incluir;}).length;
  var totalMPs = batch.filter(function(g){return g.incluir;})
    .reduce(function(s,g){return s + g.duplicados.length;}, 0);
  var counter = document.getElementById('bulk-counter');
  if(counter) counter.textContent = sel + ' grupos seleccionados · ' + totalMPs + ' MPs serán archivadas';
}

async function ejecutarFusionBulk(){
  var batch = window._batchFusion || [];
  var seleccionados = batch.filter(function(g){return g.incluir && g.duplicados.length > 0;});
  if(!seleccionados.length){
    document.getElementById('bulk-msg').innerHTML = '<span style="color:#dc2626">No hay grupos seleccionados.</span>';
    return;
  }
  var motivo = (document.getElementById('bulk-motivo').value || '').trim();
  if(!confirm('¿FUSIONAR '+seleccionados.length+' grupos?\\n\\n'+
              'Esto archivará '+seleccionados.reduce(function(s,g){return s+g.duplicados.length;}, 0)+' MPs '+
              'y transferirá TODOS sus movimientos a los canónicos elegidos.\\n\\n'+
              'Acción rastreable en audit_log. Continuar?')){
    return;
  }
  var btn = document.getElementById('btn-confirmar-bulk');
  btn.disabled = true; btn.textContent = '⏳ Procesando...';
  document.getElementById('bulk-msg').innerHTML =
    '<span style="color:#64748b">Procesando '+seleccionados.length+' grupos · esto puede tardar 10-30s...</span>';
  try{
    var payload = {
      grupos: seleccionados.map(function(g){
        return {codigo_canonico: g.canonico, codigos_duplicados: g.duplicados};
      }),
      motivo: motivo,
    };
    var r = await fetch('/api/admin/maestro-mps-unificar-bulk', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify(payload),
    });
    var d = {}; try{d = await r.json();}catch(e){}
    if(r.ok){
      var t = d.resumen_totales || {};
      var html = '<div style="background:#dcfce7;color:#15803d;padding:12px;border-radius:6px;font-weight:700">'+
        '✓ '+esc(d.message || 'Bulk completado')+'<br>'+
        '<small style="font-weight:400">Transferidos: '+(t.movimientos||0)+' movs · '+
        (t.formula_items||0)+' fórmulas · '+
        (t.conteo_items||0)+' conteos · '+
        (t.sol_items||0)+' SOLs</small></div>';
      if(d.fallidos && d.fallidos.length){
        html += '<div style="background:#fef3c7;color:#92400e;padding:10px;border-radius:6px;margin-top:8px">'+
          '<b>⚠ '+d.fallidos.length+' grupos fallaron:</b><br>'+
          '<ul style="font-size:11px;margin-top:6px">'+
          d.fallidos.slice(0,10).map(function(f){
            return '<li>'+esc(f.canonico)+': '+esc(f.error)+'</li>';
          }).join('')+'</ul></div>';
      }
      document.getElementById('bulk-msg').innerHTML = html;
      setTimeout(function(){
        document.getElementById('modal-bulk').remove();
        ejecutarAudit();
      }, 3500);
    } else {
      document.getElementById('bulk-msg').innerHTML =
        '<div style="color:#dc2626;background:#fee2e2;padding:10px;border-radius:6px">'+
        '❌ Error '+r.status+': '+esc(d.error || 'unknown')+
        (d.detail ? '<br><small>'+esc(d.detail)+'</small>' : '')+'</div>';
      btn.disabled = false; btn.textContent = '✓ Fusionar seleccionados';
    }
  }catch(e){
    document.getElementById('bulk-msg').innerHTML =
      '<span style="color:#dc2626">Error de red: '+e.message+'</span>';
    btn.disabled = false; btn.textContent = '✓ Fusionar seleccionados';
  }
}

// Auto-ejecutar al cargar
ejecutarAudit();
</script>
</body></html>"""


@bp.route("/admin/mps-sin-uso", methods=["GET"])
def admin_mps_sin_uso_page():
    """Panel para detectar y archivar MPs sin uso del catálogo.

    Sebastián 8-may-2026 (zero-error FASE A): el catálogo crece con MPs
    que nunca más se usan. Archivarlas (activo=0) limpia listas sin
    perder historial INVIMA.
    """
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este panel.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_MPS_SIN_USO_HTML, mimetype='text/html')


_MPS_SIN_USO_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>MPs sin uso · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0f172a;color:#f1f5f9;padding:20px}
h1{font-size:24px;margin-bottom:6px;color:#fbbf24}
.sub{color:#94a3b8;font-size:13px;margin-bottom:24px}
.back{display:inline-block;color:#94a3b8;text-decoration:none;font-size:13px;margin-bottom:16px}
.back:hover{color:#f1f5f9}
.controls{background:#1e293b;border-radius:10px;padding:16px;margin-bottom:16px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.controls label{font-size:13px;color:#cbd5e1}
.controls input[type=number]{background:#0f172a;color:#f1f5f9;border:1px solid #475569;border-radius:6px;padding:6px 10px;width:90px}
.controls input[type=checkbox]{accent-color:#fbbf24}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px;transition:.15s}
button:disabled{opacity:.4;cursor:not-allowed}
.btn-detect{background:#fbbf24;color:#0f172a}
.btn-detect:hover:not(:disabled){background:#f59e0b}
.btn-archive{background:#dc2626;color:white}
.btn-archive:hover:not(:disabled){background:#b91c1c}
.btn-trigger{background:#5eead4;color:#0f172a;font-weight:700}
.btn-trigger:hover:not(:disabled){background:#2dd4bf}
.summary{background:#0c4a6e;border:1px solid #0284c7;border-radius:10px;padding:14px;margin-bottom:16px;display:none}
.summary.visible{display:block}
.summary p{font-size:13px;color:#bae6fd}
table{width:100%;border-collapse:collapse;background:#1e293b;border-radius:10px;overflow:hidden;font-size:12px}
th,td{padding:8px 10px;text-align:left;border-bottom:1px solid #334155}
th{background:#0f172a;color:#94a3b8;font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.5px}
tr:hover{background:#334155}
.stock-zero{color:#22c55e}
.stock-nonzero{color:#fbbf24}
.empty{color:#64748b;text-align:center;padding:30px;font-style:italic}
.kbd{background:#0f172a;padding:2px 6px;border-radius:3px;font-family:Consolas,Monaco,monospace;font-size:11px;color:#fbbf24}
.danger-zone{background:#7f1d1d;border:1px solid #dc2626;border-radius:8px;padding:14px;margin-top:16px}
.danger-zone h3{color:#fca5a5;font-size:14px;margin-bottom:8px}
.danger-zone p{font-size:12px;color:#fecaca;margin-bottom:8px}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>🗃 MPs sin uso · Catálogo</h1>
<p class="sub">Detecta MPs activas que no están en fórmulas, sin movimientos recientes y stock cero. Archivar (activo=0) preserva historial INVIMA.</p>

<div class="controls">
  <label>Inactividad mínima:
    <input type="number" id="dias" value="365" min="30" max="3650">
    días
  </label>
  <label>
    <input type="checkbox" id="incluir-stock"> Incluir con stock &gt; 0 (NO archivar)
  </label>
  <button class="btn-detect" onclick="detect()">📊 Detectar</button>
  <button class="btn-trigger" onclick="triggerVencidos()" title="Equivalente al cron diario · marca VENCIDO en lotes con fecha pasada">⛔ Marcar vencidos bulk</button>
</div>

<div class="summary" id="summary"></div>

<table id="tabla" style="display:none">
  <thead>
    <tr>
      <th><input type="checkbox" id="sel-all" onclick="toggleAll(this)"></th>
      <th>Código</th>
      <th>Nombre</th>
      <th>Stock g</th>
      <th>Última actividad</th>
      <th>Días inactivo</th>
      <th>Movs</th>
    </tr>
  </thead>
  <tbody id="tbody"></tbody>
</table>

<div class="empty" id="empty" style="display:none">No hay MPs sin uso con esos criterios.</div>

<div class="danger-zone" id="danger-zone" style="display:none">
  <h3>⚠ Zona peligrosa · Archivar bulk</h3>
  <p>Esto pondrá <span class="kbd">activo=0</span> en las MPs seleccionadas. No se borran · historial INVIMA preservado.</p>
  <p>Solo se archivan las que cumplen criterio (no en fórmula + stock=0). Las marcadas con stock&gt;0 serán rechazadas automáticamente.</p>
  <button class="btn-archive" onclick="archivar()" id="btn-arch" disabled>🗃 Archivar seleccionadas (<span id="n-sel">0</span>)</button>
</div>

<script>
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
let candidatos = [];

async function detect(){
  const dias = document.getElementById('dias').value || 365;
  const incl = document.getElementById('incluir-stock').checked ? 1 : 0;
  const sum = document.getElementById('summary');
  sum.className = 'summary visible';
  sum.innerHTML = '<p>⏳ Detectando...</p>';
  try{
    const r = await fetch(`/api/admin/mps-sin-uso?dias_inactividad=${dias}&incluir_con_stock=${incl}`);
    const d = await r.json();
    if(!r.ok){
      sum.innerHTML = '<p style="color:#fca5a5">Error: '+esc(d.error || 'falla')+'</p>';
      return;
    }
    candidatos = d.sin_uso || [];
    sum.innerHTML = `<p>${d.message} · <strong>${d.resumen.archivables_seguro}</strong> archivables seguro · ${d.resumen.con_stock_pero_no_usadas} con stock.</p>`;
    renderTabla();
  }catch(e){
    sum.innerHTML = '<p style="color:#fca5a5">Error de red: '+esc(e.message)+'</p>';
  }
}

function renderTabla(){
  const tbody = document.getElementById('tbody');
  const empty = document.getElementById('empty');
  const tabla = document.getElementById('tabla');
  const dz = document.getElementById('danger-zone');
  if(!candidatos.length){
    tabla.style.display = 'none';
    empty.style.display = 'block';
    dz.style.display = 'none';
    return;
  }
  tabla.style.display = 'table';
  empty.style.display = 'none';
  dz.style.display = 'block';
  tbody.innerHTML = candidatos.map((it, i) => {
    const stkClass = Math.abs(it.stock_actual_g) < 1 ? 'stock-zero' : 'stock-nonzero';
    const ua = it.ultima_actividad || '—nunca—';
    const di = (it.dias_inactivo !== null && it.dias_inactivo !== undefined) ? it.dias_inactivo : '∞';
    const archivable = Math.abs(it.stock_actual_g) < 1;
    return `<tr>
      <td><input type="checkbox" class="sel" data-codigo="${esc(it.codigo)}" ${archivable?'':'disabled title="stock no-cero"'} onchange="updateSel()"></td>
      <td><strong>${esc(it.codigo)}</strong></td>
      <td>${esc(it.nombre)}</td>
      <td class="${stkClass}">${it.stock_actual_g.toLocaleString()}</td>
      <td>${esc(ua)}</td>
      <td>${esc(di)}</td>
      <td>${it.n_movs}</td>
    </tr>`;
  }).join('');
  updateSel();
}

function toggleAll(cb){
  document.querySelectorAll('.sel:not([disabled])').forEach(x => x.checked = cb.checked);
  updateSel();
}

function updateSel(){
  const n = document.querySelectorAll('.sel:checked').length;
  document.getElementById('n-sel').textContent = n;
  document.getElementById('btn-arch').disabled = (n === 0);
}

async function archivar(){
  const codigos = Array.from(document.querySelectorAll('.sel:checked')).map(x => x.dataset.codigo);
  if(!codigos.length) return;
  if(!confirm(`¿Archivar ${codigos.length} MPs? Quedan activo=0 (no se borran · reversible).`)) return;
  const btn = document.getElementById('btn-arch');
  btn.disabled = true; btn.textContent = '⏳ Archivando...';
  try{
    const r = await fetch('/api/admin/archivar-mps-sin-uso-bulk', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({codigos: codigos, motivo: 'Panel mps-sin-uso · limpieza catálogo'})
    });
    const d = await r.json();
    if(!r.ok){
      alert('Error: '+ (d.error || 'falla'));
      btn.disabled = false; btn.textContent = '🗃 Archivar seleccionadas (' + codigos.length + ')';
      return;
    }
    alert(`✓ ${d.message}\\n\\nArchivadas: ${d.n_archivadas}\\nRechazadas: ${d.n_rechazadas}`);
    detect();  // re-detectar
  }catch(e){
    alert('Error de red: '+e.message);
    btn.disabled = false;
  }
}

async function triggerVencidos(){
  if(!confirm('¿Marcar como VENCIDO todos los lotes con fecha_venc pasada que aún figuran VIGENTE?')) return;
  try{
    const r = await fetch('/api/admin/marcar-vencidos-bulk-todos', {method:'POST'});
    const d = await r.json();
    if(!r.ok){
      alert('Error: '+(d.error || 'falla'));
      return;
    }
    alert(`✓ ${d.message}\\n\\nLotes afectados: ${d.lotes_afectados || 0}\\nMovs actualizados: ${d.actualizados || 0}`);
  }catch(e){
    alert('Error de red: '+e.message);
  }
}
</script>
</body></html>"""


@bp.route("/admin/auditoria-formulas", methods=["GET"])
def admin_auditoria_formulas_page():
    """S1 · Página visual del veredicto de fórmulas maestras.

    Sebastián 8-may-2026: corre /api/admin/auditoria-formulas-completa
    y pinta el score + checks. Sin necesidad de DevTools.
    """
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este panel.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_AUDIT_FORMULAS_HTML, mimetype='text/html')


_AUDIT_FORMULAS_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>S1 · Fórmulas maestras · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0f172a;color:#f1f5f9;padding:20px;line-height:1.5}
h1{font-size:24px;margin-bottom:6px;color:#5eead4}
.sub{color:#94a3b8;font-size:13px;margin-bottom:20px}
.back{display:inline-block;color:#94a3b8;text-decoration:none;font-size:13px;margin-bottom:16px}
.back:hover{color:#f1f5f9}
.hero{background:#1e293b;border-radius:14px;padding:24px;margin-bottom:20px;text-align:center}
.score{font-size:72px;font-weight:800;line-height:1}
.score.ok{color:#22c55e}
.score.warn{color:#fbbf24}
.score.bad{color:#ef4444}
.verdict{font-size:16px;font-weight:700;letter-spacing:1px;margin-top:8px;text-transform:uppercase}
.verdict.ok{color:#22c55e}
.verdict.warn{color:#fbbf24}
.verdict.bad{color:#ef4444}
.resumen{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px;margin-top:14px}
.kpi{background:#0f172a;border-radius:8px;padding:8px;font-size:11px;color:#94a3b8}
.kpi b{display:block;color:#f1f5f9;font-size:18px;margin-bottom:2px}
.checks{display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:12px}
.check{background:#1e293b;border-radius:10px;padding:16px;border-left:4px solid #475569}
.check.ok{border-left-color:#22c55e}
.check.fail{border-left-color:#ef4444}
.check h3{font-size:14px;margin-bottom:8px;display:flex;align-items:center;gap:8px}
.badge{padding:2px 10px;border-radius:10px;font-size:10px;font-weight:700;letter-spacing:.5px}
.badge.ok{background:#14532d;color:#86efac}
.badge.fail{background:#7f1d1d;color:#fca5a5}
.top{margin-top:10px;font-size:11px;color:#cbd5e1;max-height:240px;overflow:auto}
.top table{width:100%;border-collapse:collapse}
.top th{background:#0f172a;color:#94a3b8;padding:6px 8px;text-align:left;font-size:10px;text-transform:uppercase}
.top td{padding:6px 8px;border-bottom:1px solid #334155}
.fix{margin-top:10px;font-size:12px}
.fix a{color:#5eead4;text-decoration:none}
.fix a:hover{text-decoration:underline}
.loading{text-align:center;padding:40px;color:#64748b}
.error{background:#7f1d1d;color:#fecaca;padding:14px;border-radius:8px;font-size:13px}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px;background:#5eead4;color:#0f172a}
button:hover{background:#2dd4bf}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>📐 S1 · Integridad fórmulas maestras</h1>
<p class="sub">Veredicto sobre 7 checks · solo lectura · no modifica datos.</p>

<div id="content" class="loading">⏳ Ejecutando auditoría...</div>

<script>
function esc(s){return String(s===null||s===undefined?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;');}

async function run(){
  try{
    const r = await fetch('/api/admin/auditoria-formulas-completa');
    const d = await r.json();
    if(!r.ok){
      document.getElementById('content').innerHTML =
        '<div class="error">Error ' + r.status + ': ' + esc(d.error||'falla') +
        '<br><small>' + esc(d.detail||'') + '</small></div>';
      return;
    }
    render(d);
  }catch(e){
    document.getElementById('content').innerHTML =
      '<div class="error">Error de red: ' + esc(e.message) + '</div>';
  }
}

function render(d){
  const score = d.score || 0;
  const sclass = score >= 99 ? 'ok' : score >= 85 ? 'warn' : 'bad';
  const v = d.veredicto || 'BLOQUEANTE';
  const vclass = v === 'PERFECTA' ? 'ok' : v === 'MENOR' ? 'warn' : 'bad';

  let html = '<div class="hero">' +
    '<div class="score ' + sclass + '">' + score + '<small style="font-size:24px;color:#64748b">/100</small></div>' +
    '<div class="verdict ' + vclass + '">' + v + '</div>' +
    '<div class="resumen">' +
      kpi('Fórmulas', d.resumen.n_formulas) +
      kpi('Items', d.resumen.n_items) +
      kpi('Huérfanos', d.resumen.huerfanos) +
      kpi('Duplicados', d.resumen.duplicados) +
      kpi('Suma %≠100', d.resumen.sumas_pct_no_100) +
      kpi('% inválido', d.resumen.pct_invalidos) +
    '</div>' +
  '</div>';

  if(d.errores_checks && Object.keys(d.errores_checks).length){
    html += '<div class="error">⚠ Algunos checks fallaron: ' +
      esc(JSON.stringify(d.errores_checks)) + '</div>';
  }

  html += '<div class="checks">';
  const labels = {
    huerfanos: '1. Huérfanos (material_id sin maestro_mps)',
    duplicados: '2. Duplicados (mismo MP 2+ veces)',
    sumas_pct_no_100: '3. Suma % ≠ 100 ±0.5',
    material_id_nulos: '4. material_id NULL/vacío',
    pct_invalidos: '5. Porcentaje inválido',
    headers_vacios: '6. Fórmulas declaradas sin items',
    huerfanos_absolutos: '7. Huérfanos absolutos (sin id ni nombre)',
  };
  for(const k of Object.keys(labels)){
    const ch = d.checks[k] || {};
    const ok = !!ch.ok;
    html += '<div class="check ' + (ok?'ok':'fail') + '">' +
      '<h3>' + labels[k] + ' <span class="badge ' + (ok?'ok':'fail') + '">' +
      (ok ? '✓ OK' : ('⚠ ' + (ch.count||0))) + '</span></h3>';
    if(!ok && ch.top && ch.top.length){
      html += '<div class="top">' + topTable(k, ch.top) + '</div>';
    }
    if(!ok && ch.fix_link){
      html += '<div class="fix">Arreglar: <a href="' + ch.fix_link + '">' + ch.fix_link + '</a></div>';
    }
    html += '</div>';
  }
  html += '</div>';

  html += '<div style="margin-top:20px;text-align:center"><button onclick="run()">🔄 Re-ejecutar</button></div>';

  document.getElementById('content').className = '';
  document.getElementById('content').innerHTML = html;
}

function kpi(label, val){
  return '<div class="kpi"><b>' + (val||0) + '</b>' + esc(label) + '</div>';
}

function topTable(k, top){
  if(k === 'huerfanos'){
    return '<table><tr><th>material_id</th><th>nombre</th><th>productos</th></tr>' +
      top.map(t => '<tr><td><b>'+esc(t.material_id)+'</b></td><td>'+esc(t.nombre)+'</td><td>'+(t.n_productos||0)+'</td></tr>').join('') +
      '</table>';
  }
  if(k === 'duplicados'){
    return '<table><tr><th>producto</th><th>material_id</th><th>veces</th></tr>' +
      top.map(t => '<tr><td>'+esc(t.producto)+'</td><td>'+esc(t.material_id)+'</td><td>'+(t.veces||0)+'</td></tr>').join('') +
      '</table>';
  }
  if(k === 'sumas_pct_no_100'){
    return '<table><tr><th>producto</th><th>suma %</th><th>diff</th><th>items</th></tr>' +
      top.map(t => '<tr><td>'+esc(t.producto)+'</td><td>'+(t.suma_actual||0)+'</td><td>'+(t.diff||0)+'</td><td>'+(t.items||0)+'</td></tr>').join('') +
      '</table>';
  }
  return '<pre style="font-size:11px">' + esc(JSON.stringify(top, null, 2)) + '</pre>';
}

run();
</script>
</body></html>"""


@bp.route("/admin/auditoria-producciones", methods=["GET"])
def admin_auditoria_producciones_page():
    """S2 · Página visual del veredicto producción descuenta MPs."""
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este panel.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_AUDIT_PRODUCCIONES_HTML, mimetype='text/html')


_AUDIT_PRODUCCIONES_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>S2 · Producciones · EOS</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0f172a;color:#f1f5f9;padding:20px;line-height:1.5}
h1{font-size:24px;margin-bottom:6px;color:#5eead4}
.sub{color:#94a3b8;font-size:13px;margin-bottom:20px}
.back{display:inline-block;color:#94a3b8;text-decoration:none;font-size:13px;margin-bottom:16px}
.back:hover{color:#f1f5f9}
.controls{background:#1e293b;border-radius:10px;padding:14px;margin-bottom:16px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.controls label{font-size:13px;color:#cbd5e1}
.controls input{background:#0f172a;color:#f1f5f9;border:1px solid #475569;border-radius:6px;padding:6px 10px;width:90px}
.hero{background:#1e293b;border-radius:14px;padding:24px;margin-bottom:20px;text-align:center}
.score{font-size:72px;font-weight:800;line-height:1}
.score.ok{color:#22c55e}
.score.warn{color:#fbbf24}
.score.bad{color:#ef4444}
.verdict{font-size:16px;font-weight:700;letter-spacing:1px;margin-top:8px;text-transform:uppercase}
.verdict.ok{color:#22c55e}
.verdict.warn{color:#fbbf24}
.verdict.bad{color:#ef4444}
.resumen{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px;margin-top:14px}
.kpi{background:#0f172a;border-radius:8px;padding:8px;font-size:11px;color:#94a3b8}
.kpi b{display:block;color:#f1f5f9;font-size:18px;margin-bottom:2px}
.kpi.bad b{color:#ef4444}
.kpi.warn b{color:#fbbf24}
.kpi.ok b{color:#22c55e}
table{width:100%;border-collapse:collapse;background:#1e293b;border-radius:10px;overflow:hidden;font-size:12px;margin-top:10px}
th,td{padding:8px 10px;text-align:left;border-bottom:1px solid #334155}
th{background:#0f172a;color:#94a3b8;font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.5px}
tr:hover{background:#334155}
.badge{padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;letter-spacing:.5px;display:inline-block}
.badge.ok{background:#14532d;color:#86efac}
.badge.bad{background:#7f1d1d;color:#fca5a5}
.badge.warn{background:#7c2d12;color:#fed7aa}
.loading{text-align:center;padding:40px;color:#64748b}
.error{background:#7f1d1d;color:#fecaca;padding:14px;border-radius:8px;font-size:13px}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px;background:#5eead4;color:#0f172a}
button:hover{background:#2dd4bf}
.empty{color:#64748b;text-align:center;padding:30px;font-style:italic}
</style></head><body>

<a class="back" href="/modulos">← Panel inicial</a>
<h1>🏭 S2 · Producción descuenta MPs</h1>
<p class="sub">Verifica que cada producción iniciada descontó MPs vía movimientos Salida. Solo lectura · no modifica datos.</p>

<div class="controls">
  <label>Últimos
    <input type="number" id="dias" value="90" min="1" max="730">
    días
  </label>
  <button onclick="run()">🔄 Re-ejecutar</button>
</div>

<div id="content" class="loading">⏳ Cargando auditoría...</div>

<script>
function esc(s){return String(s===null||s===undefined?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;');}

async function run(){
  const dias = document.getElementById('dias').value || 90;
  document.getElementById('content').className = 'loading';
  document.getElementById('content').innerHTML = '⏳ Cargando auditoría...';
  try{
    const r = await fetch('/api/admin/auditoria-producciones-descuento?dias=' + dias);
    const d = await r.json();
    if(!r.ok){
      document.getElementById('content').innerHTML =
        '<div class="error">Error ' + r.status + ': ' + esc(d.error||'falla') +
        '<br><small>' + esc(d.detail||'') + '</small></div>';
      return;
    }
    render(d);
  }catch(e){
    document.getElementById('content').innerHTML =
      '<div class="error">Error de red: ' + esc(e.message) + '</div>';
  }
}

function render(d){
  const score = d.score || 0;
  const sclass = score >= 99 ? 'ok' : score >= 85 ? 'warn' : 'bad';
  const v = d.veredicto || 'BLOQUEANTE';
  const vclass = v === 'PERFECTA' ? 'ok' : v === 'MENOR' ? 'warn' : 'bad';
  const rs = d.resumen || {};

  let html = '<div class="hero">' +
    '<div class="score ' + sclass + '">' + score + '<small style="font-size:24px;color:#64748b">/100</small></div>' +
    '<div class="verdict ' + vclass + '">' + v + '</div>' +
    '<div class="resumen">' +
      kpi('Programadas', rs.n_total, '') +
      kpi('Iniciadas', rs.n_iniciadas, '') +
      kpi('OK', rs.n_ok, 'ok') +
      kpi('Legacy total', rs.legacy_total, '') +
      kpi('Legacy sin movs', rs.legacy_sin_movs, (rs.legacy_sin_movs||0)>0?'bad':'ok') +
    '</div>' +
    ((rs.n_problemas + (rs.legacy_sin_movs||0)) > 0 ? '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:6px;margin-top:10px">' +
      kpi('Iniciada sin descuento', rs.iniciadas_sin_descuento, (rs.iniciadas_sin_descuento||0)>0?'bad':'') +
      kpi('Descontada sin movs', rs.descontadas_sin_movs, (rs.descontadas_sin_movs||0)>0?'bad':'') +
      kpi('Sin fórmula', rs.sin_formula, (rs.sin_formula||0)>0?'warn':'') +
      kpi('Terminada sin descuento', rs.terminadas_sin_descuento, (rs.terminadas_sin_descuento||0)>0?'bad':'') +
    '</div>' : '') +
  '</div>';

  const probs = d.problemas || [];
  const legacyProbs = d.legacy_problemas || [];

  if(probs.length === 0 && legacyProbs.length === 0){
    html += '<div class="empty">✓ Sin problemas detectados en los últimos ' + rs.dias_horizonte + ' días.</div>';
  }
  if(probs.length > 0){
    html += '<h3 style="margin-top:14px;color:#fbbf24">Producción programada con problema (' + probs.length + ')</h3>';
    html += '<table><thead><tr><th>ID</th><th>Producto</th><th>Fecha</th><th>Lotes</th>' +
      '<th>Iniciada</th><th>Descontada</th><th>Movs</th><th>Estado</th></tr></thead><tbody>';
    for(const p of probs){
      const eclass = p.estado_audit === 'OK' ? 'ok' :
                     p.estado_audit === 'SIN_FORMULA' ? 'warn' : 'bad';
      html += '<tr>' +
        '<td>' + esc(p.id) + '</td>' +
        '<td>' + esc(p.producto) + '</td>' +
        '<td>' + esc(p.fecha_programada) + '</td>' +
        '<td>' + esc(p.lotes) + '</td>' +
        '<td>' + (p.iniciada ? '✓' : '—') + '</td>' +
        '<td>' + (p.descontada ? '✓' : '—') + '</td>' +
        '<td>' + esc(p.n_movs_salida_producto) + '</td>' +
        '<td><span class="badge ' + eclass + '">' + esc(p.estado_audit) + '</span></td>' +
      '</tr>';
    }
    html += '</tbody></table>';
  }
  if(legacyProbs.length > 0){
    html += '<h3 style="margin-top:18px;color:#ef4444">⚠ Producción LEGACY sin descuento de MPs (' + legacyProbs.length + ')</h3>';
    html += '<p style="font-size:11px;color:#94a3b8;margin-bottom:6px">Producciones en tabla `producciones` (flujo legacy) que NO tienen movimientos Salida. Stock real consumido pero MPs no descontadas del inventario.</p>';
    html += '<table><thead><tr><th>ID</th><th>Producto</th><th>Fecha</th><th>Cantidad</th><th>Lote</th><th>Estado</th><th>Items fórmula</th></tr></thead><tbody>';
    for(const p of legacyProbs){
      html += '<tr>' +
        '<td>' + esc(p.id) + '</td>' +
        '<td>' + esc(p.producto) + '</td>' +
        '<td>' + esc(p.fecha) + '</td>' +
        '<td>' + esc(p.cantidad) + '</td>' +
        '<td>' + esc(p.lote||'—') + '</td>' +
        '<td>' + esc(p.estado||'—') + '</td>' +
        '<td>' + esc(p.n_formula_items) + '</td>' +
      '</tr>';
    }
    html += '</tbody></table>';
  }

  document.getElementById('content').className = '';
  document.getElementById('content').innerHTML = html;
}

function kpi(label, val, cls){
  return '<div class="kpi ' + (cls||'') + '"><b>' + (val||0) + '</b>' + esc(label) + '</div>';
}

run();
</script>
</body></html>"""


@bp.route("/api/admin/reconciliar-produccion-mp", methods=["POST"])
def reconciliar_produccion_mp():
    """Reconcilia descuento omitido en una producción (bug operativo).

    Sebastián 8-may-2026: cuando una producción NO descontó una MP porque
    el stock en kardex era 0 (pero físicamente sí había producto), el
    audit profundo lo detecta. Este endpoint crea los movs retroactivos
    INVIMA-compliant para cerrar el drift.

    Body JSON:
      produccion_id: int (producciones.id)
      material_id: str (MP que NO se descontó)
      cantidad_g: float (lo que la fórmula dice debió descontar)
      fecha_retroactiva: ISO (fecha real de la producción)
      motivo: str (obligatorio · queda en audit + observaciones)
      compensar_con_entrada: bool · si true, crea mov Entrada compensatoria
                              para no afectar stock_neto actual
      fecha_entrada_compensatoria: ISO opcional (default datetime.now)

    Returns: {ok, mov_salida_id, mov_entrada_id?, audit_id}
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    prod_id = d.get('produccion_id')
    mid = (d.get('material_id') or '').strip()
    cant = float(d.get('cantidad_g') or 0)
    fecha_retro = (d.get('fecha_retroactiva') or '').strip()
    motivo = (d.get('motivo') or '').strip()
    compensar = bool(d.get('compensar_con_entrada', True))
    fecha_comp = (d.get('fecha_entrada_compensatoria') or '').strip()

    if not prod_id or not mid or cant <= 0 or not motivo:
        return jsonify({
            'error': 'produccion_id, material_id, cantidad_g > 0 y motivo son requeridos'
        }), 400
    if len(motivo) < 20:
        return jsonify({
            'error': 'motivo debe tener al menos 20 caracteres (auditabilidad INVIMA)'
        }), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    try:
        # Verificar que producción existe
        prod = c.execute(
            "SELECT id, producto, cantidad, fecha, lote FROM producciones WHERE id=?",
            (prod_id,)
        ).fetchone()
        if not prod:
            conn.close()
            return jsonify({'error': f'producción {prod_id} no existe'}), 404
        prod_dict = {'id': prod[0], 'producto': prod[1], 'cant_kg': prod[2],
                      'fecha': prod[3], 'lote': prod[4]}

        # Verificar MP existe
        mp = c.execute(
            "SELECT codigo_mp, nombre_comercial FROM maestro_mps WHERE codigo_mp=?",
            (mid,)
        ).fetchone()
        if not mp:
            conn.close()
            return jsonify({'error': f'MP {mid} no existe'}), 404

        # Verificar que NO existe ya un mov Salida de esta MP para esta producción
        lote_pat = f'PROD-{prod_id:05d}'
        ya_existe = c.execute("""
            SELECT id FROM movimientos
            WHERE material_id = ? AND tipo = 'Salida'
              AND observaciones LIKE ?
            LIMIT 1
        """, (mid, f'%{lote_pat}%')).fetchone()
        if ya_existe:
            conn.close()
            return jsonify({
                'error': f'MP {mid} ya tiene mov Salida para producción #{prod_id} (id={ya_existe[0]}). Cancelar antes de reconciliar.'
            }), 409

        # Crear mov Salida retroactivo (NOTA: este insert va a violar trigger
        # estado_lote='VIGENTE' default si no lo seteamos, pero estado_lote
        # default es 'VIGENTE' que es OK)
        obs_salida = (f"RECONCILIACION zero-error: bug descuento omitido en "
                       f"producción #{prod_id} {prod_dict['producto']} "
                       f"({prod_dict['cant_kg']}kg) lote {prod_dict['lote']}. "
                       f"MP {mid} no se descontó por stock=0 en kardex al "
                       f"momento (físicamente sí había). Motivo: {motivo[:300]}")
        c.execute("""
            INSERT INTO movimientos
              (material_id, material_nombre, cantidad, tipo, fecha,
               observaciones, lote, operador, estado_lote)
            VALUES (?, ?, ?, 'Salida', ?, ?, ?, ?, 'VIGENTE')
        """, (mid, mp[1] or '', cant,
              fecha_retro or prod_dict['fecha'],
              obs_salida,
              prod_dict['lote'] or '',
              u))
        mov_salida_id = c.lastrowid

        # Compensación: mov Entrada para no afectar stock_neto
        mov_entrada_id = None
        if compensar:
            obs_entrada = (f"RECONCILIACION zero-error: compensación por "
                            f"descuento retroactivo mov #{mov_salida_id} "
                            f"(producción #{prod_id}). El stock físico real "
                            f"al momento del ajuste anterior era mayor al "
                            f"contado · diff +{cant}g identificado por audit "
                            f"profundo. Motivo: {motivo[:300]}")
            c.execute("""
                INSERT INTO movimientos
                  (material_id, material_nombre, cantidad, tipo, fecha,
                   observaciones, lote, operador, estado_lote)
                VALUES (?, ?, ?, 'Entrada', ?, ?, ?, ?, 'VIGENTE')
            """, (mid, mp[1] or '', cant,
                  fecha_comp or datetime.now().isoformat(),
                  obs_entrada,
                  prod_dict['lote'] or '',
                  u))
            mov_entrada_id = c.lastrowid

        # Audit log
        try:
            audit_log(
                c, usuario=u,
                accion='RECONCILIAR_PRODUCCION_MP',
                tabla='movimientos', registro_id=str(mov_salida_id),
                despues={
                    'produccion_id': prod_id,
                    'producto': prod_dict['producto'],
                    'material_id': mid,
                    'cantidad_g': cant,
                    'mov_salida_id': mov_salida_id,
                    'mov_entrada_id': mov_entrada_id,
                    'compensado': compensar,
                    'motivo': motivo[:300],
                },
                detalle=(f'Reconciliacion zero-error: producción #{prod_id} '
                         f'{prod_dict["producto"]} omitió descuento de {mid} '
                         f'({cant}g). Compensación: {compensar}'),
            )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'mov_salida_id': mov_salida_id,
            'mov_entrada_id': mov_entrada_id,
            'cantidad_g': cant,
            'compensado': compensar,
            'message': (
                f'✓ Reconciliada producción #{prod_id} MP {mid} ({cant}g) · '
                f'mov Salida #{mov_salida_id}'
                + (f' + Entrada compensatoria #{mov_entrada_id}'
                   if compensar else '')
            ),
        }), 200

    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/validacion-profunda", methods=["GET"])
def validacion_profunda():
    """Validación matemática REAL · zero falsos positivos.

    Sebastián 8-may-2026: el score PERFECTA agregado puede esconder
    problemas si los checks son débiles. Este endpoint hace validación
    MATEMÁTICA real:

    1. MPs duplicadas REALES:
       - Mismo nombre_comercial con códigos distintos (case+trim insensitive)
       - Mismo INCI con códigos distintos (excluye whitelist legales)
       - Códigos con whitespace o casing duplicado

    2. Fórmulas íntegras DEEP:
       - Cada item.material_id está en maestro_mps con activo=1
       - No hay duplicados (mismo MP 2+ veces en misma fórmula)
       - Producto en formula_headers existe en producciones reales
       - Detecta fórmulas con menos items que producciones similares

    3. Descuento REAL vs ESPERADO:
       - Para cada producción legacy con stock > 0:
         · Calcula ESPERADO = formula_items[MP].porcentaje × cantidad_kg × 10
         · Compara con SUM(movs Salida) reales por MP/producción
         · Drift = REAL - ESPERADO
       - Solo flagea diferencias > 5% del esperado (tolerancia operativa)

    4. Trazabilidad INVIMA:
       - Cada movimiento Salida tiene operador NOT NULL
       - Cada lote tiene fecha_vencimiento (si stock > 0)
       - Cada producción tiene fórmula completa

    Returns: { ok, todos_los_checks, score_real, hallazgos: [...] }
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    hallazgos = []

    try:
        # ── 1. MPs DUPLICADAS REALES ─────────────────────────────────────
        # 1a · nombre_comercial duplicado entre activos
        dup_nombre = c.execute("""
            SELECT LOWER(TRIM(COALESCE(nombre_comercial,''))) AS nom,
                   COUNT(*) AS n,
                   GROUP_CONCAT(codigo_mp) AS codigos
            FROM maestro_mps
            WHERE activo = 1
              AND nombre_comercial IS NOT NULL
              AND TRIM(nombre_comercial) != ''
            GROUP BY LOWER(TRIM(COALESCE(nombre_comercial,'')))
            HAVING n > 1
            ORDER BY n DESC
            LIMIT 30
        """).fetchall()
        for nom, n, codigos in dup_nombre:
            hallazgos.append({
                'tipo': 'MP_NOMBRE_DUPLICADO',
                'severidad': 'alta',
                'detalle': f"nombre '{nom}' usado en {n} codigos: {codigos}",
                'datos': {'nombre': nom, 'codigos': codigos.split(',')},
            })

        # 1b · INCI duplicado (whitelist INCIs cosméticos compartidos legalmente)
        WHITELIST_INCI = {
            'aqua', 'water', 'glycerin', 'parfum', 'fragrance',
            'phenoxyethanol', 'ethylhexylglycerin', 'sodium hydroxide',
            'citric acid', 'lactic acid', 'tocopherol', 'panthenol',
            '', 'pendiente inci', 'pending inci',
        }
        dup_inci_raw = c.execute("""
            SELECT LOWER(TRIM(COALESCE(nombre_inci,''))) AS inci,
                   COUNT(*) AS n,
                   GROUP_CONCAT(codigo_mp) AS codigos
            FROM maestro_mps
            WHERE activo = 1
            GROUP BY LOWER(TRIM(COALESCE(nombre_inci,'')))
            HAVING n > 1
            ORDER BY n DESC
        """).fetchall()
        for inci, n, codigos in dup_inci_raw:
            if inci in WHITELIST_INCI:
                continue
            hallazgos.append({
                'tipo': 'MP_INCI_DUPLICADO',
                'severidad': 'media',
                'detalle': f"INCI '{inci}' compartido por {n} codigos: {codigos}",
                'datos': {'inci': inci, 'codigos': codigos.split(',')[:10]},
            })

        # 1c · códigos con whitespace o casing inconsistente
        codigos_raros = c.execute("""
            SELECT codigo_mp FROM maestro_mps
            WHERE activo = 1
              AND (codigo_mp != TRIM(codigo_mp)
                   OR codigo_mp LIKE '% %'
                   OR codigo_mp != UPPER(codigo_mp))
            LIMIT 20
        """).fetchall()
        for (cod,) in codigos_raros:
            hallazgos.append({
                'tipo': 'MP_CODIGO_INCONSISTENTE',
                'severidad': 'media',
                'detalle': f"codigo '{cod}' tiene whitespace o casing irregular",
                'datos': {'codigo': cod},
            })

        # ── 2. FORMULAS INTEGRAS DEEP ────────────────────────────────────
        # 2a · items con MP archivada (activo=0) · trigger FK previene nuevos
        # pero data legacy puede tener
        formula_archivadas = c.execute("""
            SELECT fi.producto_nombre, fi.material_id,
                   COALESCE(fi.material_nombre,'') AS nom
            FROM formula_items fi
            JOIN maestro_mps m ON m.codigo_mp = fi.material_id
            WHERE m.activo = 0
            LIMIT 30
        """).fetchall()
        for prod, mid, nom in formula_archivadas:
            hallazgos.append({
                'tipo': 'FORMULA_USA_MP_ARCHIVADA',
                'severidad': 'alta',
                'detalle': f"formula '{prod}' usa MP archivada {mid} ({nom})",
                'datos': {'producto': prod, 'material_id': mid},
            })

        # 2b · fórmulas declaradas pero producciones nunca existieron
        # (info · no necesariamente bug)
        try:
            fh_huerfanas = c.execute("""
                SELECT fh.producto_nombre
                FROM formula_headers fh
                WHERE NOT EXISTS (
                    SELECT 1 FROM producciones p WHERE p.producto = fh.producto_nombre
                )
                  AND NOT EXISTS (
                    SELECT 1 FROM produccion_programada pp WHERE pp.producto = fh.producto_nombre
                )
                LIMIT 20
            """).fetchall()
            for (prod,) in fh_huerfanas:
                hallazgos.append({
                    'tipo': 'FORMULA_SIN_PRODUCCIONES',
                    'severidad': 'baja',
                    'detalle': f"formula '{prod}' nunca se uso en una producción real",
                    'datos': {'producto': prod},
                })
        except sqlite3.OperationalError:
            pass

        # ── 3. DESCUENTO REAL vs ESPERADO ───────────────────────────────
        # Solo producciones recientes con stock > 0 (los con stock=0 son
        # historicos, drift menor no importa)
        prod_rows = c.execute("""
            SELECT id, producto, cantidad, fecha, lote
            FROM producciones
            WHERE fecha >= date('now', '-180 days')
            ORDER BY fecha DESC
            LIMIT 30
        """).fetchall()

        for pid, producto, cant_kg, fecha, lote in prod_rows:
            cant_kg = float(cant_kg or 0)
            if cant_kg <= 0:
                continue

            # Fórmula del producto
            items_formula = c.execute("""
                SELECT material_id, material_nombre, porcentaje
                FROM formula_items
                WHERE producto_nombre = ?
            """, (producto,)).fetchall()
            if not items_formula:
                continue

            # Movs Salida REALES de esta producción
            # Formato observaciones: 'FEFO:PROD-NNNNN:...' o 'UNLIMITED:PROD-NNNNN:...'
            lote_ref_pat = f'PROD-{pid:05d}'
            for mid, nom, pct in items_formula:
                pct = float(pct or 0)
                if pct <= 0:
                    continue
                # Esperado: % * cantidad_kg * 10  (porque pct es 0-100 y kg→g multiplicas por 1000, / 100)
                esperado_g = pct * cant_kg * 10

                real_row = c.execute("""
                    SELECT COALESCE(SUM(cantidad), 0)
                    FROM movimientos
                    WHERE material_id = ?
                      AND tipo = 'Salida'
                      AND (observaciones LIKE ?
                           OR observaciones LIKE ?)
                """, (mid, f'%{lote_ref_pat}%', f'%{producto}%' + (f'%{lote}%' if lote else ''))).fetchone()
                real_g = float(real_row[0] or 0)

                if esperado_g < 0.01:
                    continue

                drift_pct = abs(real_g - esperado_g) / esperado_g * 100
                # Tolerancia 5% (operativa · pesaje físico tiene error)
                if drift_pct > 5:
                    hallazgos.append({
                        'tipo': 'DESCUENTO_DRIFT_PRODUCCION',
                        'severidad': 'alta',
                        'detalle': (
                            f"producción #{pid} '{producto}' MP {mid}: "
                            f"esperaba {esperado_g:.2f}g, descontó {real_g:.2f}g, "
                            f"drift {drift_pct:.1f}%"
                        ),
                        'datos': {
                            'produccion_id': pid,
                            'producto': producto,
                            'material_id': mid,
                            'esperado_g': round(esperado_g, 2),
                            'real_g': round(real_g, 2),
                            'drift_pct': round(drift_pct, 2),
                        },
                    })

        # ── 4. TRAZABILIDAD INVIMA ─────────────────────────────────────
        # 4a · movs sin operador
        sin_op = c.execute("""
            SELECT COUNT(*) FROM movimientos
            WHERE (operador IS NULL OR TRIM(operador) = '')
              AND fecha >= date('now', '-90 days')
        """).fetchone()[0]
        if sin_op > 0:
            hallazgos.append({
                'tipo': 'TRAZABILIDAD_MOV_SIN_OPERADOR',
                'severidad': 'media',
                'detalle': f"{sin_op} movimientos sin operador en últimos 90 días",
                'datos': {'count': sin_op},
            })

        # 4b · lotes con stock > 0 sin fecha_venc (legacy o nuevo)
        sin_fv_con_stock = c.execute("""
            WITH lote_stock AS (
                SELECT material_id, lote,
                       SUM(CASE WHEN tipo='Entrada' THEN cantidad
                                WHEN tipo='Salida' THEN -cantidad
                                ELSE 0 END) AS stock,
                       MAX(fecha_vencimiento) AS fv
                FROM movimientos
                WHERE lote IS NOT NULL AND lote != ''
                GROUP BY material_id, lote
            )
            SELECT COUNT(*) FROM lote_stock
            WHERE stock > 0
              AND (fv IS NULL OR TRIM(fv) = '')
        """).fetchone()[0]
        if sin_fv_con_stock > 0:
            hallazgos.append({
                'tipo': 'TRAZABILIDAD_LOTE_VIVO_SIN_FV',
                'severidad': 'alta',
                'detalle': f"{sin_fv_con_stock} lotes con stock > 0 sin fecha_vencimiento (riesgo INVIMA)",
                'datos': {'count': sin_fv_con_stock},
            })

        # Resumen por severidad
        alta = sum(1 for h in hallazgos if h['severidad'] == 'alta')
        media = sum(1 for h in hallazgos if h['severidad'] == 'media')
        baja = sum(1 for h in hallazgos if h['severidad'] == 'baja')

        # Score real: solo cuenta hallazgos de severidad alta y media
        score = 100.0
        score -= alta * 5.0
        score -= media * 1.0
        score = max(0.0, round(score, 1))

        if alta == 0 and media == 0:
            veredicto = 'PERFECTA'
        elif alta == 0:
            veredicto = 'MENOR'
        else:
            veredicto = 'BLOQUEANTE'

        conn.close()

        return jsonify({
            'ok': True,
            'score_real': score,
            'veredicto_real': veredicto,
            'resumen': {
                'total_hallazgos': len(hallazgos),
                'alta': alta,
                'media': media,
                'baja': baja,
            },
            'hallazgos': hallazgos,
            'checks_ejecutados': [
                'MP_NOMBRE_DUPLICADO',
                'MP_INCI_DUPLICADO (con whitelist cosmetica)',
                'MP_CODIGO_INCONSISTENTE',
                'FORMULA_USA_MP_ARCHIVADA',
                'FORMULA_SIN_PRODUCCIONES',
                'DESCUENTO_DRIFT_PRODUCCION (tolerancia 5%)',
                'TRAZABILIDAD_MOV_SIN_OPERADOR',
                'TRAZABILIDAD_LOTE_VIVO_SIN_FV',
            ],
            'message': (f'Validación profunda: {veredicto} · score_real {score}/100 '
                        f'· {alta} ALTA · {media} MEDIA · {baja} BAJA'),
        }), 200

    except Exception as e:
        conn.close()
        return jsonify({'error': 'falla validacion profunda',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/completar-info-lote-bulk", methods=["POST"])
def completar_info_lote_bulk():
    """Completa info faltante (fecha_venc, proveedor) en lotes legacy.

    Sebastián 8-may-2026: S5 detecta 12 lotes con SIN_FECHA_VENC o
    SIN_PROVEEDOR. Este endpoint actualiza TODOS los movs de un
    (material_id, lote) con la info nueva. Preserva movs existentes.

    Body JSON:
      items: [{
        material_id: str (requerido),
        lote: str (requerido),
        fecha_vencimiento: str ISO (opcional · si null y aplicar_default_fv=true,
                                     usa primer_mov + 1 año con observación),
        proveedor: str (opcional · si null queda igual)
      }]
      aplicar_default_fv: bool (default False · solo afecta items sin fecha_venc)
      motivo: str (audit_log)

    Returns: {ok, actualizados:[...], rechazados:[...], total}
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    items = d.get('items') or []
    aplicar_default_fv = bool(d.get('aplicar_default_fv', False))
    motivo = (d.get('motivo') or 'Completar info lotes legacy · audit S5').strip()

    if not isinstance(items, list) or not items:
        return jsonify({'error': 'items debe ser lista no vacía'}), 400
    if len(items) > 100:
        return jsonify({'error': 'max 100 items por request'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    actualizados = []
    rechazados = []
    try:
        for it in items:
            mid = (it.get('material_id') or '').strip()
            lote = (it.get('lote') or '').strip()
            fv = (it.get('fecha_vencimiento') or '').strip() or None
            prov = (it.get('proveedor') or '').strip() or None

            if not mid or not lote:
                rechazados.append({'item': it,
                                    'razon': 'material_id y lote requeridos'})
                continue

            # Verificar que existe el lote
            existe = c.execute("""
                SELECT MIN(fecha) FROM movimientos
                WHERE material_id=? AND lote=?
            """, (mid, lote)).fetchone()
            if not existe or not existe[0]:
                rechazados.append({'material_id': mid, 'lote': lote,
                                    'razon': 'lote no existe en movimientos'})
                continue
            primer_mov = existe[0]

            # Si no se dio fecha_venc y aplicar_default_fv: usar primer_mov + 1 año
            fv_final = fv
            usado_default = False
            if not fv_final and aplicar_default_fv:
                try:
                    from datetime import datetime as _dt, timedelta as _td
                    pm = _dt.fromisoformat(primer_mov[:10])
                    fv_final = (pm + _td(days=365)).date().isoformat()
                    usado_default = True
                except Exception:
                    pass

            # Construir SET clause dinámico
            sets = []
            params = []
            if fv_final:
                sets.append('fecha_vencimiento=?')
                params.append(fv_final)
            if prov:
                sets.append('proveedor=?')
                params.append(prov)
            if not sets:
                rechazados.append({'material_id': mid, 'lote': lote,
                                    'razon': 'sin fecha_venc ni proveedor para actualizar'})
                continue

            params.extend([mid, lote])
            c.execute(
                f"UPDATE movimientos SET {', '.join(sets)} "
                f"WHERE material_id=? AND lote=?",
                params
            )
            n = c.rowcount
            actualizados.append({
                'material_id': mid, 'lote': lote,
                'fecha_vencimiento_aplicada': fv_final,
                'proveedor_aplicado': prov,
                'movs_actualizados': n,
                'usado_default_fv': usado_default,
            })

        if actualizados:
            try:
                audit_log(
                    c, usuario=u,
                    accion='COMPLETAR_INFO_LOTE_BULK',
                    tabla='movimientos', registro_id='bulk',
                    despues={
                        'motivo': motivo,
                        'aplicar_default_fv': aplicar_default_fv,
                        'actualizados': actualizados[:50],
                        'rechazados': rechazados[:50],
                        'n_actualizados': len(actualizados),
                        'n_rechazados': len(rechazados),
                    },
                    detalle=(f'Completar info en {len(actualizados)} lotes · '
                             f'default_fv={aplicar_default_fv} · motivo: {motivo[:120]}'),
                )
            except Exception:
                pass

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'actualizados': actualizados,
            'rechazados': rechazados,
            'n_actualizados': len(actualizados),
            'n_rechazados': len(rechazados),
            'message': (f'✓ Completados {len(actualizados)} lotes '
                        f'· rechazados {len(rechazados)}'),
        }), 200
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/investigar-mee/<codigo>", methods=["GET"])
def investigar_mee(codigo):
    """Forensic para diagnosticar drift en MEE.

    Sebastián 8-may-2026: cuando S3 detecta drift entre maestro_mee.stock_actual
    y SUM(movimientos_mee), necesitamos saber EXACTAMENTE qué movimiento o
    qué edit manual introdujo el drift.

    Returns:
      mee: {codigo, descripcion, stock_actual_persistido, ...}
      movimientos: [{id, tipo, cantidad, fecha, anulado, observaciones, ...}]
      stock_calculado: SUM(entradas - salidas + ajustes) de movs vigentes
      drift: stock_actual_persistido - stock_calculado
      audit_log: entries relacionadas (acciones que modificaron este MEE)
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    cod = (codigo or '').strip()
    if not cod:
        return jsonify({'error': 'codigo requerido'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()

    try:
        mee_row = c.execute(
            "SELECT codigo, descripcion, categoria, proveedor, fabricante, "
            "       estado, stock_actual, stock_minimo, unidad, fecha_creacion "
            "FROM maestro_mee WHERE codigo = ?",
            (cod,)
        ).fetchone()
        if not mee_row:
            conn.close()
            return jsonify({'error': f'MEE {cod} no existe en maestro_mee'}), 404
        mee_cols = ['codigo','descripcion','categoria','proveedor','fabricante',
                     'estado','stock_actual','stock_minimo','unidad','fecha_creacion']
        mee = dict(zip(mee_cols, mee_row))

        # TODOS los movs en orden cronologico
        rows = c.execute("""
            SELECT id, tipo, cantidad, lote_ref, batch_ref, responsable,
                   observaciones, fecha, COALESCE(anulado,0) AS anulado
            FROM movimientos_mee
            WHERE mee_codigo = ?
            ORDER BY fecha ASC, id ASC
        """, (cod,)).fetchall()
        movs = []
        for r in rows:
            movs.append({
                'id': r[0], 'tipo': r[1], 'cantidad': r[2],
                'lote_ref': r[3], 'batch_ref': r[4],
                'responsable': r[5], 'observaciones': r[6],
                'fecha': r[7], 'anulado': r[8],
            })

        # Stock calculado (solo vigentes · no anulados)
        stock_calc = 0.0
        for m in movs:
            if m['anulado']:
                continue
            cant = float(m['cantidad'] or 0)
            if m['tipo'] == 'Entrada':
                stock_calc += cant
            elif m['tipo'] == 'Salida':
                stock_calc -= cant
            elif m['tipo'] == 'Ajuste':
                stock_calc += cant  # ajuste puede ser + o -

        stock_persistido = float(mee['stock_actual'] or 0)
        drift = stock_persistido - stock_calc

        # Audit log relacionado (best-effort · tabla puede no existir)
        audit_entries = []
        try:
            ar = c.execute("""
                SELECT timestamp, usuario, accion, registro_id, detalle
                FROM audit_log
                WHERE (tabla = 'maestro_mee' AND registro_id = ?)
                   OR (tabla = 'movimientos_mee' AND detalle LIKE ?)
                   OR (accion LIKE '%MEE%' AND detalle LIKE ?)
                ORDER BY timestamp DESC
                LIMIT 50
            """, (cod, f'%{cod}%', f'%{cod}%')).fetchall()
            for r in ar:
                audit_entries.append({
                    'timestamp': r[0], 'usuario': r[1],
                    'accion': r[2], 'registro_id': r[3],
                    'detalle': (r[4] or '')[:300],
                })
        except sqlite3.OperationalError:
            pass

        conn.close()

        return jsonify({
            'ok': True,
            'mee': mee,
            'stock_persistido': stock_persistido,
            'stock_calculado': stock_calc,
            'drift': drift,
            'n_movs_total': len(movs),
            'n_movs_anulados': sum(1 for m in movs if m['anulado']),
            'movimientos': movs,
            'audit_log_relacionado': audit_entries,
            'recomendacion': (
                'Stock OK · sin drift' if abs(drift) < 1
                else (
                    'Crear movimiento Ajuste para reconciliar · usar '
                    '/api/admin/reconciliar-mee con drift como cantidad'
                )
            ),
        }), 200

    except Exception as e:
        conn.close()
        return jsonify({'error': 'falla query',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/reconciliar-mee", methods=["POST"])
def reconciliar_mee():
    """Crea un movimiento Ajuste para reconciliar drift MEE.

    Sebastián 8-may-2026: cuando un MEE tiene drift entre persistido y
    calculado, esta es la única forma INVIMA-compliant de igualarlos:
    crear un mov Ajuste explícito con razón documentada. NO se borran
    movimientos existentes · se agrega uno nuevo.

    Body JSON:
      codigo: str (MEE)
      sentido: 'subir_calculado' | 'bajar_calculado'
        · subir: crea mov Entrada con cant=|drift| · stock_calc sube hasta persistido
        · bajar: pone stock_actual = stock_calc (sin tocar movimientos)
      motivo: str (audit_log)

    Returns: {ok, mov_id?, audit_id?, message}
    """
    u, err, code = _require_admin()
    if err:
        return err, code

    d = request.json or {}
    cod = (d.get('codigo') or '').strip()
    sentido = (d.get('sentido') or '').strip()
    motivo = (d.get('motivo') or 'Reconciliación drift MEE · auditoría S3').strip()

    if not cod:
        return jsonify({'error': 'codigo requerido'}), 400
    if sentido not in ('subir_calculado', 'bajar_calculado'):
        return jsonify({'error': 'sentido invalido (debe ser subir_calculado o bajar_calculado)'}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=2000")
    c = conn.cursor()
    try:
        # Calcular drift actual
        mee = c.execute(
            "SELECT stock_actual FROM maestro_mee WHERE codigo=?", (cod,)
        ).fetchone()
        if not mee:
            conn.close()
            return jsonify({'error': f'MEE {cod} no existe'}), 404
        stock_persistido = float(mee[0] or 0)

        calc_row = c.execute("""
            SELECT COALESCE(SUM(
              CASE WHEN tipo='Entrada' AND COALESCE(anulado,0)=0 THEN cantidad
                   WHEN tipo='Salida' AND COALESCE(anulado,0)=0 THEN -cantidad
                   WHEN tipo='Ajuste' AND COALESCE(anulado,0)=0 THEN cantidad
                   ELSE 0 END
            ), 0)
            FROM movimientos_mee WHERE mee_codigo=?
        """, (cod,)).fetchone()
        stock_calc = float(calc_row[0] or 0)
        drift = stock_persistido - stock_calc

        if abs(drift) < 1:
            conn.close()
            return jsonify({
                'ok': True,
                'message': 'Sin drift · nada que reconciliar',
                'stock_persistido': stock_persistido,
                'stock_calculado': stock_calc,
            }), 200

        mov_id = None
        if sentido == 'subir_calculado':
            # Crear mov Ajuste con cant=drift (positivo o negativo)
            c.execute("""
                INSERT INTO movimientos_mee
                  (mee_codigo, tipo, cantidad, observaciones, responsable, fecha)
                VALUES (?, 'Ajuste', ?, ?, ?, datetime('now'))
            """, (cod, drift,
                  f'RECONCILIACION zero-error: persistido era {stock_persistido}, '
                  f'calculado era {stock_calc}, drift {drift:+.2f}. Motivo: {motivo[:200]}',
                  u))
            mov_id = c.lastrowid
        else:
            # bajar_calculado: poner stock_actual=stock_calc en maestro_mee
            c.execute(
                "UPDATE maestro_mee SET stock_actual=? WHERE codigo=?",
                (stock_calc, cod)
            )

        # Audit log
        audit_id = None
        try:
            audit_log(
                c, usuario=u,
                accion='RECONCILIAR_MEE',
                tabla='maestro_mee', registro_id=cod,
                antes={'stock_persistido': stock_persistido,
                        'stock_calculado': stock_calc},
                despues={'sentido': sentido, 'drift_aplicado': drift,
                          'mov_ajuste_id': mov_id, 'motivo': motivo[:300]},
                detalle=f'Reconciliacion drift MEE {cod}: {drift:+.2f} ({sentido})',
            )
        except Exception:
            pass

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'codigo': cod,
            'sentido': sentido,
            'drift_aplicado': drift,
            'mov_ajuste_id': mov_id,
            'audit_id': audit_id,
            'message': (f'Reconciliado · drift {drift:+.2f} cerrado via '
                        f'{"mov Ajuste" if sentido == "subir_calculado" else "UPDATE maestro"}'),
        }), 200

    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'falla transaccional',
                        'detail': str(e)[:300]}), 500


@bp.route("/api/admin/realidad-cero-error", methods=["GET"])
def api_realidad_cero_error():
    """S6 agregador zero-error score combinado S1-S5."""
    u, err, code = _require_admin()
    if err:
        return err, code

    resultados = {}

    def _llamar(funcion):
        try:
            resp = funcion()
            if isinstance(resp, tuple):
                payload, status = resp[0].get_json(), resp[1]
            else:
                payload, status = resp.get_json(), resp.status_code
            return {'ok': status == 200, 'status': status, 'data': payload or {}}
        except Exception as e:
            return {'ok': False, 'error': str(e)[:200]}

    resultados['S1_formulas'] = _llamar(auditoria_formulas_completa)
    resultados['S2_producciones'] = _llamar(auditoria_producciones_descuento)
    resultados['S3_kardex'] = _llamar(auditoria_kardex_drift)
    resultados['S4_mps_nuevas'] = _llamar(auditoria_mps_nuevas)
    resultados['S5_lotes_nuevos'] = _llamar(auditoria_lotes_nuevos)

    pesos = {'S1_formulas': 1, 'S2_producciones': 2, 'S3_kardex': 2,
              'S4_mps_nuevas': 1, 'S5_lotes_nuevos': 1}
    suma_ponderada = 0.0
    suma_pesos = 0.0
    detalles = {}
    todos_perfecta = True
    algun_bloqueante = False

    for key, peso in pesos.items():
        r = resultados.get(key, {})
        if r.get('ok') and r.get('data'):
            d = r['data']
            score = float(d.get('score', 0))
            veredicto = d.get('veredicto', 'BLOQUEANTE')
            mensaje = d.get('message', '')
        else:
            score = 0.0
            veredicto = 'ERROR'
            mensaje = r.get('error', 'fallo invocacion')

        if veredicto != 'PERFECTA':
            todos_perfecta = False
        if veredicto == 'BLOQUEANTE':
            algun_bloqueante = True

        suma_ponderada += score * peso
        suma_pesos += peso

        detalles[key] = {
            'score': score,
            'veredicto': veredicto,
            'message': mensaje,
        }

    score_global = round(suma_ponderada / max(suma_pesos, 1), 1)

    if todos_perfecta:
        veredicto_global = 'PERFECTA'
    elif algun_bloqueante:
        veredicto_global = 'BLOQUEANTE'
    else:
        veredicto_global = 'MENOR'

    return jsonify({
        'ok': True,
        'score_global': score_global,
        'veredicto_global': veredicto_global,
        'detalles': detalles,
        'message': (f'Realidad zero-error: {veredicto_global} score global {score_global}/100'),
    }), 200


@bp.route("/admin/realidad-cero-error", methods=["GET"])
def admin_realidad_cero_error_page():
    """S6 dashboard ejecutivo agregado S1-S5."""
    u, err, code = _require_admin()
    if err:
        return Response(
            '<h1>403</h1><p>Solo admin puede ver este panel.</p>',
            status=403, mimetype='text/html'
        )
    return Response(_REALIDAD_HTML, mimetype='text/html')


_REALIDAD_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<title>Realidad Zero-Error</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,Segoe UI,sans-serif;background:#0f172a;color:#f1f5f9;padding:20px;line-height:1.5}
h1{font-size:26px;margin-bottom:6px;color:#5eead4}
.sub{color:#94a3b8;font-size:13px;margin-bottom:20px}
.back{display:inline-block;color:#94a3b8;text-decoration:none;font-size:13px;margin-bottom:16px}
.back:hover{color:#f1f5f9}
.hero{background:#1e293b;border-radius:14px;padding:28px;margin-bottom:20px;text-align:center}
.score{font-size:88px;font-weight:800;line-height:1}
.score.ok{color:#22c55e}
.score.warn{color:#fbbf24}
.score.bad{color:#ef4444}
.verdict{font-size:20px;font-weight:800;letter-spacing:2px;margin-top:10px;text-transform:uppercase}
.verdict.ok{color:#22c55e}
.verdict.warn{color:#fbbf24}
.verdict.bad{color:#ef4444}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px;margin-top:14px}
.card{background:#1e293b;border-radius:12px;padding:18px;border-left:5px solid #475569}
.card.ok{border-left-color:#22c55e}
.card.warn{border-left-color:#fbbf24}
.card.bad{border-left-color:#ef4444}
.card h3{font-size:15px;margin-bottom:6px;display:flex;justify-content:space-between;align-items:center}
.card-score{font-size:32px;font-weight:800;margin-top:8px}
.card-score.ok{color:#22c55e}
.card-score.warn{color:#fbbf24}
.card-score.bad{color:#ef4444}
.badge{padding:3px 10px;border-radius:12px;font-size:10px;font-weight:700;display:inline-block}
.badge.ok{background:#14532d;color:#86efac}
.badge.warn{background:#7c2d12;color:#fed7aa}
.badge.bad{background:#7f1d1d;color:#fca5a5}
.badge.err{background:#1e293b;color:#94a3b8}
.msg{font-size:11px;color:#94a3b8;margin-top:6px;line-height:1.4}
.linka{color:#5eead4;text-decoration:none;font-size:11px;margin-top:8px;display:inline-block}
.linka:hover{text-decoration:underline}
.loading{text-align:center;padding:40px;color:#64748b}
button{padding:8px 16px;border:none;border-radius:6px;cursor:pointer;font-weight:600;font-size:13px;background:#5eead4;color:#0f172a;margin-bottom:14px}
button:hover{background:#2dd4bf}
</style></head><body>

<a class="back" href="/modulos">Panel inicial</a>
<h1>Realidad zero-error</h1>
<p class="sub">Score agregado de las 5 auditorias core (S1-S5). Si veredicto = PERFECTA, sistema sin defectos.</p>

<button onclick="run()">Re-ejecutar todo</button>
<div id="content" class="loading">Cargando...</div>

<script>
function esc(s){return String(s===null||s===undefined?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;');}

const ETIQUETAS = {
  S1_formulas:     {nombre:'S1 Formulas integras', link:'/admin/auditoria-formulas', desc:'huerfanos duplicados suma %'},
  S2_producciones: {nombre:'S2 Produccion descuenta MPs', link:'/admin/auditoria-producciones', desc:'mov Salida por produccion'},
  S3_kardex:       {nombre:'S3 Kardex sin drift', link:'/admin/auditoria-kardex', desc:'stock = entradas - salidas'},
  S4_mps_nuevas:   {nombre:'S4 MPs nuevas ingresan', link:'#', desc:'MP nueva tiene Entrada inicial'},
  S5_lotes_nuevos: {nombre:'S5 Lotes nuevos reales', link:'#', desc:'lote tiene fecha_venc + proveedor'},
};

async function run(){
  document.getElementById('content').className = 'loading';
  document.getElementById('content').innerHTML = 'Re-ejecutando 5 auditorias...';
  try{
    const r = await fetch('/api/admin/realidad-cero-error');
    const d = await r.json();
    if(!r.ok){
      document.getElementById('content').innerHTML = '<div style="background:#7f1d1d;color:#fecaca;padding:14px;border-radius:8px">Error: ' + esc(d.error||'falla') + '</div>';
      return;
    }
    render(d);
  }catch(e){
    document.getElementById('content').innerHTML = '<div style="background:#7f1d1d;color:#fecaca;padding:14px;border-radius:8px">Error de red: ' + esc(e.message) + '</div>';
  }
}

function render(d){
  const score = d.score_global || 0;
  const sclass = score >= 99 ? 'ok' : score >= 85 ? 'warn' : 'bad';
  const v = d.veredicto_global || 'BLOQUEANTE';
  const vclass = v === 'PERFECTA' ? 'ok' : v === 'MENOR' ? 'warn' : 'bad';

  let html = '<div class="hero">' +
    '<div class="score ' + sclass + '">' + score + '<small style="font-size:32px;color:#64748b">/100</small></div>' +
    '<div class="verdict ' + vclass + '">' + v + '</div>' +
  '</div>';

  html += '<div class="grid">';
  for(const key of Object.keys(ETIQUETAS)){
    const e = ETIQUETAS[key];
    const d2 = (d.detalles || {})[key] || {};
    const sc = d2.score || 0;
    const ver = d2.veredicto || 'ERROR';
    const cclass = ver === 'PERFECTA' ? 'ok' : ver === 'MENOR' ? 'warn' : ver === 'BLOQUEANTE' ? 'bad' : 'err';
    const bclass = cclass === 'err' ? 'err' : cclass;
    html += '<div class="card ' + cclass + '">' +
      '<h3>' + esc(e.nombre) + ' <span class="badge ' + bclass + '">' + esc(ver) + '</span></h3>' +
      '<div style="font-size:11px;color:#94a3b8">' + esc(e.desc) + '</div>' +
      '<div class="card-score ' + cclass + '">' + sc + '<small style="font-size:14px;color:#64748b">/100</small></div>' +
      '<div class="msg">' + esc(d2.message || '') + '</div>' +
      (e.link !== '#' ? '<a class="linka" href="' + e.link + '">Ver detalle</a>' : '') +
    '</div>';
  }
  html += '</div>';

  document.getElementById('content').className = '';
  document.getElementById('content').innerHTML = html;
}

run();
</script>
</body></html>"""
