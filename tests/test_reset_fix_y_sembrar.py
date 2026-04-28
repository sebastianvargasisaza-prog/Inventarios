"""Tests del fix de reset+replay (compensación FEFO) y endpoint sembrar maestro."""
import io
import os
import sqlite3

import openpyxl
from openpyxl.styles import PatternFill

from .conftest import TEST_PASSWORD, csrf_headers


GREEN = PatternFill('solid', fgColor='FF92D050')


def _login(app, user="sebastian"):
    c = app.test_client()
    r = c.post(
        "/login",
        data={"username": user, "password": TEST_PASSWORD},
        headers=csrf_headers(),
        follow_redirects=False,
    )
    assert r.status_code == 302
    return c


def _excel_minimo(verdes, no_verdes=()):
    """Genera xlsx en memoria con header en fila 5, datos desde fila 6.
    verdes: list of (codigo, nombre, proveedor, lote, cant_g)
    no_verdes: igual pero sin verde."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'INVENTARIO'
    ws['A1'] = 'INVENTARIO DE MATERIAS PRIMAS'
    ws['A5'] = 'CÓDIGO MP'
    ws['B5'] = 'NOMBRE INCI'
    ws['C5'] = 'NOMBRE COMERCIAL'
    ws['E5'] = 'PROVEEDOR'
    ws['G5'] = 'N° LOTE'
    ws['H5'] = 'CANT. CONTEO(g)'
    ws['J5'] = 'ESTANTERÍA'
    ws['K5'] = 'POS.'
    r = 6
    for cod, nombre, prov, lote, cant in verdes:
        ws.cell(row=r, column=1, value=cod).fill = GREEN
        ws.cell(row=r, column=2, value=nombre)
        ws.cell(row=r, column=3, value=nombre)
        ws.cell(row=r, column=5, value=prov)
        ws.cell(row=r, column=7, value=lote)
        ws.cell(row=r, column=8, value=cant)
        r += 1
    for cod, nombre, prov, lote, cant in no_verdes:
        ws.cell(row=r, column=1, value=cod)
        ws.cell(row=r, column=2, value=nombre)
        ws.cell(row=r, column=3, value=nombre)
        ws.cell(row=r, column=5, value=prov)
        ws.cell(row=r, column=7, value=lote)
        ws.cell(row=r, column=8, value=cant)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_sembrar_maestro_requiere_admin(app, db_clean):
    c = app.test_client()
    buf = _excel_minimo([('MP_TEST_X', 'Test X', 'Inchemical', 'L1', 100)])
    r = c.post(
        "/api/admin/sembrar-maestro-desde-excel",
        data={'file': (buf, 'test.xlsx')},
        content_type='multipart/form-data',
        headers=csrf_headers(),
    )
    assert r.status_code == 401


def test_sembrar_maestro_inserta_nuevos(app, db_clean):
    """Si el código del Excel no está en maestro_mps, se inserta."""
    c = _login(app)
    # Asegurar que MP_NEW_SEED no esté
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute("DELETE FROM maestro_mps WHERE codigo_mp='MP_NEW_SEED'")
    conn.commit()
    conn.close()

    buf = _excel_minimo([
        ('MP_NEW_SEED', 'Test Seed', 'Agenquimicos', 'L_SEED_1', 500),
    ])
    r = c.post(
        "/api/admin/sembrar-maestro-desde-excel",
        data={'file': (buf, 'test.xlsx')},
        content_type='multipart/form-data',
        headers=csrf_headers(),
    )
    assert r.status_code == 200, r.get_json()
    d = r.get_json()
    assert d['ok'] is True
    assert d['nuevos_count'] >= 1
    # Verificar en BD
    conn = sqlite3.connect(os.environ["DB_PATH"])
    row = conn.execute(
        "SELECT codigo_mp, proveedor FROM maestro_mps WHERE codigo_mp='MP_NEW_SEED'"
    ).fetchone()
    conn.close()
    assert row is not None
    assert row[1] == 'Agenquimicos'


def test_sembrar_maestro_actualiza_proveedor_vacio(app, db_clean):
    """Si MP existe pero sin proveedor, lo actualiza desde Excel."""
    c = _login(app)
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute(
        "INSERT OR REPLACE INTO maestro_mps (codigo_mp, nombre_comercial, proveedor, activo) "
        "VALUES (?,?,?,?)",
        ("MP_NO_PROV", "MP sin proveedor", "", 1),
    )
    conn.commit()
    conn.close()

    buf = _excel_minimo([
        ('MP_NO_PROV', 'MP sin proveedor', 'Lyphar', 'L_X', 50),
    ])
    r = c.post(
        "/api/admin/sembrar-maestro-desde-excel",
        data={'file': (buf, 'test.xlsx')},
        content_type='multipart/form-data',
        headers=csrf_headers(),
    )
    assert r.status_code == 200
    d = r.get_json()
    assert d['proveedores_actualizados_count'] >= 1

    conn = sqlite3.connect(os.environ["DB_PATH"])
    prov = conn.execute(
        "SELECT proveedor FROM maestro_mps WHERE codigo_mp='MP_NO_PROV'"
    ).fetchone()[0]
    conn.close()
    assert prov == 'Lyphar'


def test_sembrar_maestro_no_toca_movimientos(app, db_clean):
    """No debe insertar nada en `movimientos`."""
    c = _login(app)
    conn = sqlite3.connect(os.environ["DB_PATH"])
    movs_pre = conn.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
    conn.close()

    buf = _excel_minimo([
        ('MP_NO_MOV_X', 'NoMov', 'Inchemical', 'LX', 999),
    ])
    r = c.post(
        "/api/admin/sembrar-maestro-desde-excel",
        data={'file': (buf, 'test.xlsx')},
        content_type='multipart/form-data',
        headers=csrf_headers(),
    )
    assert r.status_code == 200

    conn = sqlite3.connect(os.environ["DB_PATH"])
    movs_post = conn.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
    conn.close()
    assert movs_post == movs_pre  # Cero movimientos creados


def test_reset_preview_compensa_lotes_con_salidas(app, db_clean):
    """Preview debe mostrar lotes_compensados_por_salidas_post si los hay."""
    c = _login(app)
    # Insertar movimiento de salida FEFO sobre un lote que estará en Excel verde
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute("DELETE FROM movimientos WHERE material_id='MP_RST_TEST'")
    # Salida FEFO post-día-cero
    conn.execute(
        "INSERT INTO movimientos (material_id, material_nombre, cantidad, tipo, fecha, lote, observaciones, operador) "
        "VALUES (?,?,?,?,?,?,?,?)",
        ('MP_RST_TEST', 'Reset Test', 200.0, 'Salida', '2026-04-20',
         'L_RST_1', 'FEFO: produccion test', 'test'),
    )
    conn.commit()
    conn.close()

    # Excel verde dice: queda 100g de ese lote
    buf = _excel_minimo([
        ('MP_RST_TEST', 'Reset Test', 'Inchemical', 'L_RST_1', 100),
    ])
    r = c.post(
        "/api/admin/inventario-reset-preview",
        data={'file': (buf, 'test.xlsx')},
        content_type='multipart/form-data',
        headers=csrf_headers(),
    )
    assert r.status_code == 200, r.get_json()
    d = r.get_json()
    iniciales = d['plan']['entradas_iniciales_a_crear']
    # Debe reportar que hay al menos 1 lote compensado
    assert iniciales['lotes_compensados_por_salidas_post_count'] >= 1
    # Debe haber sumado las salidas (100 + 200 = 300)
    sample = iniciales['lotes_compensados_sample_top10']
    assert any(
        s['codigo_mp'] == 'MP_RST_TEST' and s['cantidad_inicial_dia_cero_g'] == 300.0
        for s in sample
    )

    # Limpiar
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute("DELETE FROM movimientos WHERE material_id='MP_RST_TEST'")
    conn.commit()
    conn.close()
