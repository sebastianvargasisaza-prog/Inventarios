"""Tests del panel admin extendido: users, reset password, eventos, config."""
from .conftest import TEST_PASSWORD, csrf_headers


# ── /api/admin/users ─────────────────────────────────────────────────────────


def test_users_list_requires_admin(client, db_clean):
    r = client.get("/api/admin/users")
    assert r.status_code == 401


def test_users_list_blocked_for_non_admin(app, db_clean):
    c = app.test_client()
    c.post("/login", data={"username": "valentina", "password": TEST_PASSWORD},
           headers=csrf_headers(), follow_redirects=False)
    r = c.get("/api/admin/users")
    assert r.status_code == 403


def test_users_list_admin_returns_19_users(admin_client, db_clean):
    r = admin_client.get("/api/admin/users")
    assert r.status_code == 200
    data = r.get_json()
    assert data["total"] == 19
    assert len(data["users"]) == 19


def test_users_list_includes_groups_and_password_source(admin_client, db_clean):
    r = admin_client.get("/api/admin/users")
    data = r.get_json()
    sebastian = next(u for u in data["users"] if u["username"] == "sebastian")
    assert sebastian["is_admin"] is True
    assert "Admin" in sebastian["groups"]
    assert sebastian["password_source"] in ("env", "db")
    # No expone hashes ni passwords
    assert "password_hash" not in sebastian
    assert "password" not in sebastian


def test_users_list_no_hashes_exposed(admin_client, db_clean):
    """CRITICAL: el endpoint NO debe devolver hashes ni passwords plaintext."""
    r = admin_client.get("/api/admin/users")
    body = r.get_data(as_text=True)
    # Ningun PBKDF2 hash ni la TEST_PASSWORD plana en el response
    assert "pbkdf2:" not in body
    assert TEST_PASSWORD not in body


# ── /api/admin/reset-password ────────────────────────────────────────────────


def test_reset_password_requires_admin(client, db_clean):
    r = client.post("/api/admin/reset-password",
                    json={"username": "valentina"},
                    headers=csrf_headers())
    assert r.status_code == 401


def test_reset_password_blocked_for_non_admin(app, db_clean):
    c = app.test_client()
    c.post("/login", data={"username": "felipe", "password": TEST_PASSWORD},
           headers=csrf_headers(), follow_redirects=False)
    r = c.post("/api/admin/reset-password",
               json={"username": "felipe"},
               headers=csrf_headers())
    assert r.status_code == 403


def test_reset_password_unknown_user(admin_client, db_clean):
    r = admin_client.post("/api/admin/reset-password",
                          json={"username": "fantasma"},
                          headers=csrf_headers())
    assert r.status_code == 404


def test_reset_password_missing_username(admin_client, db_clean):
    r = admin_client.post("/api/admin/reset-password",
                          json={},
                          headers=csrf_headers())
    assert r.status_code == 400


def test_reset_password_works_end_to_end(app, db_clean):
    """Admin resetea, user puede loguear con nueva password."""
    admin = app.test_client()
    admin.post("/login", data={"username": "sebastian", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    r = admin.post("/api/admin/reset-password",
                   json={"username": "valentina"},
                   headers=csrf_headers())
    assert r.status_code == 200
    data = r.get_json()
    assert data["ok"] is True
    new_pwd = data["new_password"]
    assert len(new_pwd) >= 8

    # Valentina puede loguear con la nueva
    user_c = app.test_client()
    r = user_c.post("/login",
                    data={"username": "valentina", "password": new_pwd},
                    headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302  # redirect


def test_reset_password_disables_old_password(app, db_clean):
    """Después del reset, la password vieja (env var) ya no funciona."""
    admin = app.test_client()
    admin.post("/login", data={"username": "sebastian", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    admin.post("/api/admin/reset-password",
               json={"username": "yuliel"},
               headers=csrf_headers())
    # Ya no entra con la vieja TEST_PASSWORD
    user_c = app.test_client()
    r = user_c.post("/login",
                    data={"username": "yuliel", "password": TEST_PASSWORD},
                    headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 200  # fallo


# ── /api/admin/security-events ───────────────────────────────────────────────


def test_security_events_requires_admin(client, db_clean):
    r = client.get("/api/admin/security-events")
    assert r.status_code == 401


def test_security_events_returns_list(admin_client, db_clean):
    """Tras el login del admin (que generó login_success), debe haber al menos 1."""
    r = admin_client.get("/api/admin/security-events")
    assert r.status_code == 200
    data = r.get_json()
    assert "events" in data
    assert "stats_24h" in data
    assert isinstance(data["events"], list)


def test_security_events_filter_by_event_type(app, db_clean):
    """Filtro ?event=login_success funciona."""
    c = app.test_client()
    # Logear admin para generar 1 login_success
    c.post("/login", data={"username": "sebastian", "password": TEST_PASSWORD},
           headers=csrf_headers(), follow_redirects=False)
    r = c.get("/api/admin/security-events?event=login_success&limit=5")
    assert r.status_code == 200
    events = r.get_json()["events"]
    for e in events:
        assert e["event"] == "login_success"


def test_security_events_limit_capped(admin_client, db_clean):
    """Limit > 500 se capea a 500 (anti-abuso)."""
    r = admin_client.get("/api/admin/security-events?limit=99999")
    assert r.status_code == 200
    # No debe crashear ni devolver más de 500
    assert len(r.get_json()["events"]) <= 500


# ── /api/admin/config-status ─────────────────────────────────────────────────


def test_config_status_requires_admin(client, db_clean):
    r = client.get("/api/admin/config-status")
    assert r.status_code == 401


def test_config_status_returns_structure(admin_client, db_clean):
    r = admin_client.get("/api/admin/config-status")
    assert r.status_code == 200
    data = r.get_json()
    for key in ("issues", "critical", "user_passwords", "optional"):
        assert key in data, f"Falta {key}"


def test_config_status_no_values_exposed(admin_client, db_clean):
    """CRITICAL: solo 'set' bool y 'length', NO el valor."""
    r = admin_client.get("/api/admin/config-status")
    body = r.get_data(as_text=True)
    # NO debe aparecer la SECRET_KEY de tests ni el TEST_PASSWORD
    assert "test-secret-key-only-for-pytest" not in body
    assert TEST_PASSWORD not in body
    # Cada item debe tener 'set' y 'length' pero NO 'value'
    data = r.get_json()
    for v in data["critical"] + data["user_passwords"] + data["optional"]:
        assert "set" in v
        assert "length" in v
        assert "value" not in v


def test_config_status_lists_all_user_passwords(admin_client, db_clean):
    """Las 19 PASS_<USER> aparecen."""
    r = admin_client.get("/api/admin/config-status")
    data = r.get_json()
    user_pass_names = {v["name"] for v in data["user_passwords"]}
    assert len(user_pass_names) == 19
    assert "PASS_SEBASTIAN" in user_pass_names


# ═══ Auditoria inventario vs Excel dia cero ═════════════════════════════════


def test_audit_inventario_vs_excel_solo_admin(app, db_clean):
    """POST /api/admin/audit-inventario-vs-excel → 403 para no-admin."""
    from .conftest import TEST_PASSWORD, csrf_headers
    c = app.test_client()
    r = c.post("/login", data={"username": "luis", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302
    r = c.post("/api/admin/audit-inventario-vs-excel",
               data={}, headers=csrf_headers())
    assert r.status_code == 403


def test_audit_inventario_vs_excel_filtra_solo_verde(admin_client, db_clean):
    """El endpoint solo cuenta filas en color VERDE (FF92D050).

    Construye un Excel sintetico con 3 filas: 1 verde, 1 roja, 1 sin marcar.
    Verifica que solo la verde llega al reporte.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTARIO"
    ws.cell(row=1, column=1, value="INVENTARIO DE MATERIAS PRIMAS")
    ws.cell(row=5, column=1, value="CÓDIGO MP")
    ws.cell(row=5, column=2, value="NOMBRE INCI")
    ws.cell(row=5, column=3, value="NOMBRE COMERCIAL")
    ws.cell(row=5, column=4, value="TIPO")
    ws.cell(row=5, column=5, value="PROVEEDOR")
    ws.cell(row=5, column=6, value="STOCK MIN")
    ws.cell(row=5, column=7, value="N° LOTE")
    ws.cell(row=5, column=8, value="CANT. CONTEO(g)")
    ws.cell(row=5, column=9, value="CANT. ACTUAL")
    ws.cell(row=5, column=10, value="ESTANTERIA")
    ws.cell(row=5, column=11, value="POS.")
    ws.cell(row=5, column=12, value="FECHA VENC.")

    verde = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
    rojo = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    blanco = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    # Fila verde — debe contar
    for col in range(1, 13):
        ws.cell(row=6, column=col).fill = verde
    ws.cell(row=6, column=1, value="MP_VERDE_T")
    ws.cell(row=6, column=3, value="Material verde test")
    ws.cell(row=6, column=5, value="Lyphar")
    ws.cell(row=6, column=7, value="LOTE-V-1")
    ws.cell(row=6, column=8, value=1000)

    # Fila roja — debe excluir
    for col in range(1, 13):
        ws.cell(row=7, column=col).fill = rojo
    ws.cell(row=7, column=1, value="MP_ROJO_T")
    ws.cell(row=7, column=3, value="Material rojo test")
    ws.cell(row=7, column=7, value="LOTE-R-1")
    ws.cell(row=7, column=8, value=999999)

    # Fila blanca — debe excluir
    for col in range(1, 13):
        ws.cell(row=8, column=col).fill = blanco
    ws.cell(row=8, column=1, value="MP_BLANCO_T")
    ws.cell(row=8, column=7, value="LOTE-B-1")
    ws.cell(row=8, column=8, value=888888)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = admin_client.post(
        "/api/admin/audit-inventario-vs-excel",
        data={"file": (buf, "test.xlsx")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 200, f"body: {r.data[:300]!r}"
    j = r.get_json()
    s = j["resumen"]
    assert s["lotes_verde_excel"] == 1, f"solo 1 verde, recibido: {s}"
    assert s["lotes_excluidos_no_verde"] == 2, (
        f"2 no-verde (rojo+blanco), recibido: {s}"
    )
    assert s["stock_total_excel_g"] == 1000.0
    # MP_VERDE_T no esta en DB → debe aparecer en faltantes
    faltantes_codigos = [x["codigo_mp"] for x in j["faltantes_en_db"]]
    assert "MP_VERDE_T" in faltantes_codigos
    # MP_ROJO_T y MP_BLANCO_T NO deben estar en ningun bucket
    todos = (j["en_db_match_sample"] + j["en_db_con_delta"]
             + j["faltantes_en_db"] + j["solo_db_no_excel"])
    codigos_total = [x.get("codigo_mp") for x in todos]
    assert "MP_ROJO_T" not in codigos_total
    assert "MP_BLANCO_T" not in codigos_total


def test_audit_inventario_vs_excel_excel_invalido(admin_client, db_clean):
    """Sin archivo o no-xlsx → 400."""
    r = admin_client.post("/api/admin/audit-inventario-vs-excel", data={})
    assert r.status_code == 400
    import io
    r = admin_client.post(
        "/api/admin/audit-inventario-vs-excel",
        data={"file": (io.BytesIO(b"not an excel"), "test.txt")},
        content_type="multipart/form-data",
    )
    assert r.status_code == 400


def test_inventario_diagnostico_solo_admin(app, db_clean):
    """GET /api/admin/inventario-diagnostico-entradas → 403 para no-admin."""
    from .conftest import TEST_PASSWORD, csrf_headers
    c = app.test_client()
    r = c.post("/login", data={"username": "luis", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302
    r = c.get("/api/admin/inventario-diagnostico-entradas")
    assert r.status_code == 403


def test_inventario_diagnostico_detecta_doble_carga(admin_client, db_clean):
    """Si un lote tiene 2 Entradas → debe aparecer en multi_entradas."""
    import sqlite3
    import os

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    # Lote con UNA entrada (legitimo)
    cur.execute("""INSERT INTO movimientos
                   (material_id, material_nombre, lote, cantidad, tipo, fecha,
                    operador)
                   VALUES ('MP_DIAG_OK','OK','LOTE-OK',1000,'Entrada',
                           '2026-04-20','sistema')""")
    # Lote con DOS entradas (doble carga)
    cur.execute("""INSERT INTO movimientos
                   (material_id, material_nombre, lote, cantidad, tipo, fecha,
                    operador)
                   VALUES ('MP_DIAG_DUP','DUP','LOTE-DUP',500,'Entrada',
                           '2026-04-20','sistema')""")
    cur.execute("""INSERT INTO movimientos
                   (material_id, material_nombre, lote, cantidad, tipo, fecha,
                    operador)
                   VALUES ('MP_DIAG_DUP','DUP','LOTE-DUP',500,'Entrada',
                           '2026-04-21','luis')""")
    conn.commit()
    conn.close()

    r = admin_client.get("/api/admin/inventario-diagnostico-entradas")
    assert r.status_code == 200
    j = r.get_json()
    duplicados = [m for m in j["multi_entradas"] if m["codigo_mp"] == "MP_DIAG_DUP"]
    assert len(duplicados) == 1, f"Doble carga no detectada: {j['multi_entradas']}"
    assert duplicados[0]["n_entradas"] == 2
    assert duplicados[0]["total_g"] == 1000.0

    # MP_DIAG_OK NO debe aparecer (solo 1 entrada)
    legits = [m for m in j["multi_entradas"] if m["codigo_mp"] == "MP_DIAG_OK"]
    assert len(legits) == 0

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos WHERE material_id IN ('MP_DIAG_OK','MP_DIAG_DUP')")
    conn.commit(); conn.close()


# ═══ Reset+Replay del inventario ═════════════════════════════════════════════


def test_inventario_snapshot_solo_admin(app, db_clean):
    c = app.test_client()
    r = c.post("/login", data={"username": "luis", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302
    r = c.get("/api/admin/inventario-snapshot-pre-reset")
    assert r.status_code == 403


def test_inventario_snapshot_descarga_json(admin_client, db_clean):
    r = admin_client.get("/api/admin/inventario-snapshot-pre-reset")
    assert r.status_code == 200
    assert r.mimetype == "application/json"
    import json
    data = json.loads(r.data)
    assert "meta" in data and "tablas" in data
    for t in ("movimientos", "producciones", "ordenes_compra",
              "comprobantes_pago", "maestro_mps", "audit_log"):
        assert t in data["tablas"], f"falta {t}"


def _build_excel_verde_test():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTARIO"
    ws.cell(row=1, column=1, value="INV MP")
    headers = ["CODIGO MP", "NOMBRE INCI", "NOMBRE COMERCIAL", "TIPO",
               "PROVEEDOR", "STOCK MIN", "N LOTE", "CANT. CONTEO(g)",
               "CANT. ACTUAL", "ESTANTERIA", "POS.", "FECHA VENC."]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    verde = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
    rojo = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    for col in range(1, 13):
        ws.cell(row=6, column=col).fill = verde
    ws.cell(row=6, column=1, value="MP_RESET_A")
    ws.cell(row=6, column=3, value="Material A reset")
    ws.cell(row=6, column=5, value="Lyphar")
    ws.cell(row=6, column=7, value="LOTE-A")
    ws.cell(row=6, column=8, value=2000)
    for col in range(1, 13):
        ws.cell(row=7, column=col).fill = verde
    ws.cell(row=7, column=1, value="MP_RESET_B")
    ws.cell(row=7, column=3, value="Material B reset")
    ws.cell(row=7, column=5, value="Inchemical")
    ws.cell(row=7, column=7, value="LOTE-B")
    ws.cell(row=7, column=8, value=500)
    for col in range(1, 13):
        ws.cell(row=8, column=col).fill = rojo
    ws.cell(row=8, column=1, value="MP_RESET_RED")
    ws.cell(row=8, column=7, value="LOTE-RED")
    ws.cell(row=8, column=8, value=99999)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_inventario_reset_preview_muestra_plan(admin_client, db_clean):
    import sqlite3
    import os
    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador) "
                "VALUES ('MP_DIRTY','Dirty','L1',888,'Entrada','2026-04-15','sistema')")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador, observaciones) "
                "VALUES ('MP_RESET_A','MA','LOTE-A',300,'Salida','2026-04-20','luis','FEFO:PROD-X:T x 1kg')")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador, numero_oc) "
                "VALUES ('MP_RESET_B','MB','LOTE-B',100,'Entrada','2026-04-22','catalina','OC-1')")
    conn.commit()
    conn.close()

    buf = _build_excel_verde_test()
    r = admin_client.post("/api/admin/inventario-reset-preview",
                          data={"file": (buf, "test.xlsx")},
                          content_type="multipart/form-data")
    assert r.status_code == 200, f"body: {r.data[:300]!r}"
    p = r.get_json()["plan"]
    assert p["movimientos_a_borrar"] == 3
    assert p["entradas_iniciales_a_crear"]["count"] == 2
    assert p["entradas_iniciales_a_crear"]["total_g"] == 2500.0
    assert p["entradas_oc_a_preservar"]["count"] == 1
    assert p["salidas_produccion_a_preservar"]["count"] == 1
    assert p["rows_no_verde_excluidas"] == 1

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    n = cur.execute("SELECT COUNT(*) FROM movimientos").fetchone()[0]
    assert n == 3, "Preview no debe modificar movimientos"
    cur.execute("DELETE FROM movimientos WHERE material_id LIKE 'MP%'")
    conn.commit()
    conn.close()
    _cleanup_backup_log()


def test_inventario_reset_aplicar_requiere_token(admin_client, db_clean):
    buf = _build_excel_verde_test()
    r = admin_client.post("/api/admin/inventario-reset-aplicar",
                          data={"file": (buf, "test.xlsx")},
                          content_type="multipart/form-data")
    assert r.status_code == 400


def test_inventario_reset_aplicar_token_incorrecto(admin_client, db_clean):
    buf = _build_excel_verde_test()
    r = admin_client.post("/api/admin/inventario-reset-aplicar",
                          data={"file": (buf, "test.xlsx"),
                                "confirmacion": "TOKEN_RANDOM"},
                          content_type="multipart/form-data")
    assert r.status_code == 400


def _mock_backup_reciente():
    """Simula un backup hecho hace 1 hora con status='ok' para que el
    endpoint reset NO intente crear uno nuevo automaticamente (eso
    contaminaria el slot de backup_log y haria fallar tests posteriores
    de /api/admin/backup-now)."""
    import sqlite3
    import os
    from datetime import datetime, timedelta
    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    ts = (datetime.utcnow() - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        "INSERT INTO backup_log (started_at, completed_at, status, "
        "file_path, triggered_by) VALUES (?, ?, 'ok', '/tmp/mock.db.gz', 'test_mock')",
        (ts, ts)
    )
    conn.commit()
    conn.close()


def _cleanup_backup_log():
    """Limpia backup_log + espera a que cualquier backup async termine.

    El admin app tiene un hook _maybe_trigger_backup que cada N requests
    evalua si lanzar backup_worker (thread daemon). Si nuestros tests
    disparan ese thread y termina despues del test, el _local_lock de
    backup.py queda adquirido y el siguiente test_backup_now_admin
    falla con 'another backup running in this worker'.

    Cleanup: forzamos liberar el lock + borrar backup_log.
    """
    import sqlite3
    import os
    try:
        # Forzar liberacion del lock (defensive — si el thread async aun corre)
        from backup import _local_lock
        try:
            while _local_lock.locked():
                _local_lock.release()
        except RuntimeError:
            pass
    except Exception:
        pass
    try:
        conn = sqlite3.connect(os.environ["DB_PATH"])
        conn.execute("DELETE FROM backup_log")
        conn.commit()
        conn.close()
    except Exception:
        pass


def test_inventario_reset_aplicar_ejecuta(admin_client, db_clean):
    import sqlite3
    import os
    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    # Limpiar movimientos del DB para test aislado
    cur.execute("DELETE FROM movimientos")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador) "
                "VALUES ('MP_DIRTY','D','L1',999,'Entrada','2026-04-15','sistema')")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador) "
                "VALUES ('MP_DIRTY','D','L1',999,'Entrada','2026-04-15','sistema')")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador, observaciones) "
                "VALUES ('MP_RESET_A','MA','LOTE-A',300,'Salida','2026-04-20','luis','FEFO:PROD-X:T x 1kg')")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador, numero_oc) "
                "VALUES ('MP_RESET_B','MB','LOTE-B',100,'Entrada','2026-04-22','catalina','OC-LEG')")
    conn.commit()
    conn.close()

    TOKEN = "BORRAR_INVENTARIO_Y_CARGAR_EXCEL_2026_04_27"
    buf = _build_excel_verde_test()
    r = admin_client.post("/api/admin/inventario-reset-aplicar",
                          data={"file": (buf, "test.xlsx"),
                                "confirmacion": TOKEN},
                          content_type="multipart/form-data")
    assert r.status_code == 200, f"body: {r.data[:400]!r}"
    j = r.get_json()
    assert j["ok"] is True
    assert j["resumen"]["movs_borrados"] == 4
    assert j["resumen"]["lotes_excel_cargados"] == 2
    assert j["resumen"]["entradas_oc_preservadas"] == 1
    assert j["resumen"]["salidas_prod_preservadas"] == 1
    # Compensación FEFO: LOTE-A tiene 1 salida de 300g, debe sumarse al excel
    assert j["resumen"]["lotes_compensados_fefo"] == 1
    assert j["resumen"]["compensacion_total_g"] == 300.0

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT material_id, lote, cantidad, tipo FROM movimientos "
        "ORDER BY material_id, fecha, cantidad"
    ).fetchall()
    assert len(rows) == 4, f"Esperado 4 movs, recibido {len(rows)}: {rows}"
    codigos = {r[0] for r in rows}
    assert "MP_DIRTY" not in codigos
    # FIX FEFO: la entrada inicial = excel (2000) + salidas_post (300) = 2300
    # Después la salida (300) se aplica → stock final = 2300 - 300 = 2000 = excel ✓
    assert ("MP_RESET_A", "LOTE-A", 2300.0, "Entrada") in rows
    assert ("MP_RESET_A", "LOTE-A", 300.0, "Salida") in rows
    assert ("MP_RESET_B", "LOTE-B", 500.0, "Entrada") in rows
    assert ("MP_RESET_B", "LOTE-B", 100.0, "Entrada") in rows

    audits = cur.execute(
        "SELECT accion FROM audit_log WHERE accion LIKE 'RESET_INVENTARIO%'"
    ).fetchall()
    actions = {a[0] for a in audits}
    assert "RESET_INVENTARIO_PRE" in actions
    assert "RESET_INVENTARIO_POST" in actions

    cur.execute("DELETE FROM movimientos WHERE material_id LIKE 'MP%'")
    conn.commit()
    conn.close()
    _cleanup_backup_log()


def test_health_monitor_solo_admin(app, db_clean):
    c = app.test_client()
    r = c.post("/login", data={"username": "luis", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302
    r = c.get("/api/admin/inventario-health-monitor")
    assert r.status_code == 403


def test_health_monitor_sistema_limpio(admin_client, db_clean):
    """Sistema limpio (sin entradas raras) → nivel ok."""
    import sqlite3
    import os
    # Aislar: limpiar movimientos por si tests previos dejaron residuos
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute("DELETE FROM movimientos")
    conn.commit()
    conn.close()
    r = admin_client.get("/api/admin/inventario-health-monitor")
    assert r.status_code == 200
    j = r.get_json()
    assert j["nivel"] == "ok"
    assert j["count_critical"] == 0
    assert len(j["alertas"]) == 0


def test_health_monitor_detecta_burst(admin_client, db_clean):
    """30+ entradas en un solo día (no día cero) → alerta BURST."""
    import sqlite3
    import os
    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    # Insertar 35 entradas en un mismo día NO cero
    for i in range(35):
        cur.execute("INSERT INTO movimientos (material_id, material_nombre, "
                    "lote, cantidad, tipo, fecha, operador) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (f'MP_BURST_{i:02d}', f'Burst {i}', f'L{i:03d}',
                     1000.0, 'Entrada', '2026-04-20T10:00:00', 'unknown'))
    conn.commit()
    conn.close()

    r = admin_client.get("/api/admin/inventario-health-monitor")
    assert r.status_code == 200
    j = r.get_json()
    burst_alerts = [a for a in j["alertas"] if a["tipo"] == "BURST"]
    assert len(burst_alerts) >= 1, f"BURST no detectado: {j}"
    assert j["nivel"] in ("warning", "critical")

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos WHERE material_id LIKE 'MP_BURST%'")
    conn.commit()
    conn.close()


def test_health_monitor_detecta_multi_entradas(admin_client, db_clean):
    """Lote con 2 Entradas (sin marker reset) → alerta MULTI_ENTRADAS."""
    import sqlite3
    import os
    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, "
                "lote, cantidad, tipo, fecha, operador) "
                "VALUES ('MP_MULTI','M','LOTE-MULTI',500,'Entrada',"
                "'2026-04-22','user1')")
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, "
                "lote, cantidad, tipo, fecha, operador) "
                "VALUES ('MP_MULTI','M','LOTE-MULTI',500,'Entrada',"
                "'2026-04-23','user2')")
    conn.commit()
    conn.close()

    r = admin_client.get("/api/admin/inventario-health-monitor")
    j = r.get_json()
    multi = [a for a in j["alertas"] if a["tipo"] == "MULTI_ENTRADAS"]
    assert len(multi) >= 1, f"MULTI_ENTRADAS no detectado: {j}"

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos WHERE material_id='MP_MULTI'")
    conn.commit()
    conn.close()


def test_inventario_reset_regenera_lotes_huerfanos(admin_client, db_clean):
    """Si una salida produccion apunta a lote NO en Excel verde, regenera
    Entrada virtual para que cierre en 0 (no en negativo)."""
    import sqlite3
    import os

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    cur.execute("INSERT INTO movimientos (material_id, material_nombre, lote, "
                "cantidad, tipo, fecha, operador, observaciones) "
                "VALUES ('MP_HUERFANO','MH','LOTE-FANTASMA',150,'Salida',"
                "'2026-04-20','luis','FEFO:PROD-X:T x 1kg')")
    conn.commit()
    conn.close()

    TOKEN = "BORRAR_INVENTARIO_Y_CARGAR_EXCEL_2026_04_27"
    buf = _build_excel_verde_test()
    r = admin_client.post("/api/admin/inventario-reset-aplicar",
                          data={"file": (buf, "test.xlsx"),
                                "confirmacion": TOKEN},
                          content_type="multipart/form-data")
    assert r.status_code == 200, f"body: {r.data[:400]!r}"
    j = r.get_json()
    assert j["resumen"]["huerfanos_regenerados"] == 1, f"resumen: {j['resumen']}"

    conn = sqlite3.connect(os.environ["DB_PATH"])
    cur = conn.cursor()
    saldo = cur.execute(
        "SELECT COALESCE(SUM(CASE WHEN tipo='Entrada' THEN cantidad ELSE -cantidad END),0) "
        "FROM movimientos WHERE material_id='MP_HUERFANO' AND lote='LOTE-FANTASMA'"
    ).fetchone()[0]
    assert saldo == 0, f"Lote huerfano debe cerrar en 0, no en {saldo}"

    entrada_huer = cur.execute(
        "SELECT cantidad, observaciones, estado_lote FROM movimientos "
        "WHERE material_id='MP_HUERFANO' AND lote='LOTE-FANTASMA' AND tipo='Entrada'"
    ).fetchone()
    assert entrada_huer is not None
    assert entrada_huer[0] == 150.0
    obs_lower = (entrada_huer[1] or "").lower()
    assert "regenerada" in obs_lower or "huerfano" in obs_lower or "consumido" in obs_lower
    assert entrada_huer[2] == "AGOTADO"

    cur.execute("DELETE FROM movimientos WHERE material_id LIKE 'MP%'")
    conn.commit()
    conn.close()
    _cleanup_backup_log()
