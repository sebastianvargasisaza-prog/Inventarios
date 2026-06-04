"""Cruce Maestro · Excel (CÓD. BATCH canónico) ↔ maestro/fórmulas/inventario.

Cubre: reporte read-only detecta INCI vacío / código fuera de fórmula, y
?aplicar=inci rellena el INCI vacío del maestro desde el Excel (sin sobrescribir).
"""
import io
import os
import sqlite3
from .conftest import TEST_PASSWORD, csrf_headers


def _login(app, user="sebastian"):
    c = app.test_client()
    r = c.post("/login", data={"username": user, "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302, r.data
    return c


def _excel_maestro(rows, producto="PROD CRUCE T1"):
    """Construye un Excel maestro de 1 hoja: header en fila 4, MPs debajo.
    rows = [(inci, comercial, cod, pct), ...]"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = producto
    ws.cell(row=1, column=1, value=producto)
    ws.cell(row=4, column=1, value="#")
    ws.cell(row=4, column=2, value="NOMBRE INCI")
    ws.cell(row=4, column=3, value="NOMBRE COMERCIAL")
    ws.cell(row=4, column=4, value="CÓD. BATCH")
    ws.cell(row=4, column=5, value="% FÓRMULA")
    for i, (inci, com, cod, pct) in enumerate(rows, start=1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=inci)
        ws.cell(row=r, column=3, value=com)
        ws.cell(row=r, column=4, value=cod)
        ws.cell(row=r, column=5, value=pct)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_cruce_maestro_reporte_y_backfill_inci(app, db_clean):
    c = _login(app)
    # maestro: una MP con INCI vacío (se debe rellenar), otra con INCI ok
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp, nombre_inci, tipo_material, activo) "
                 "VALUES ('MP-CR-A', '', 'MP', 1)")  # INCI vacío
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp, nombre_inci, tipo_material, activo) "
                 "VALUES ('MP-CR-B', 'Glycerin', 'MP', 1)")
    conn.commit(); conn.close()

    xls = _excel_maestro([
        ("BORON NITRIDE", "Boron nitride", "MP-CR-A", 0.02),
        ("GLYCERIN", "Glicerina", "MP-CR-B", 5.0),
    ], producto="PROD CRUCE T1")

    # 1) read-only: detecta INCI_VACIO en MP-CR-A
    r = c.post("/api/admin/cruce-maestro",
               data={"file": (io.BytesIO(xls), "m.xlsx")},
               headers=csrf_headers(),
               content_type="multipart/form-data")
    assert r.status_code == 200, r.data
    d = r.get_json()
    assert d["resumen"]["inci_vacio"] >= 1, d["resumen"]
    assert d["aplicado"] is False
    # INCI no se tocó aún
    conn = sqlite3.connect(os.environ["DB_PATH"])
    inci_a = conn.execute("SELECT nombre_inci FROM maestro_mps WHERE codigo_mp='MP-CR-A'").fetchone()[0]
    conn.close()
    assert (inci_a or "") == "", "read-only no debe modificar"

    # 2) aplicar=inci: rellena el vacío desde el Excel, NO toca el que ya tenía
    r2 = c.post("/api/admin/cruce-maestro?aplicar=inci",
                data={"file": (io.BytesIO(xls), "m.xlsx")},
                headers=csrf_headers(),
                content_type="multipart/form-data")
    assert r2.status_code == 200, r2.data
    assert r2.get_json()["aplicado"] is True
    assert r2.get_json()["resumen"]["inci_rellenados"] >= 1, r2.get_json()["resumen"]
    conn = sqlite3.connect(os.environ["DB_PATH"])
    a2 = conn.execute("SELECT nombre_inci FROM maestro_mps WHERE codigo_mp='MP-CR-A'").fetchone()[0]
    b2 = conn.execute("SELECT nombre_inci FROM maestro_mps WHERE codigo_mp='MP-CR-B'").fetchone()[0]
    conn.close()
    assert a2 == "BORON NITRIDE", f"INCI debe rellenarse · {a2}"
    assert b2 == "Glycerin", "INCI existente no debe sobrescribirse"


def test_cruce_maestro_detecta_codigos_huerfanos(app, db_clean):
    """Un código usado en formula_items que NO existe en maestro_mps se reporta
    como huérfano (fórmula no producible · código muerto/legacy)."""
    c = _login(app)
    conn = sqlite3.connect(os.environ["DB_PATH"])
    # el trigger FK impide insertar formula_items con código inexistente; los
    # huérfanos son data legacy (código existía, luego desapareció del maestro).
    # Se simula: crear código → insertar fórmula → borrar el código del maestro.
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp, nombre_inci, tipo_material, activo) "
                 "VALUES ('MPDEADCODE99', 'Muerto', 'MP', 1)")
    conn.execute("INSERT INTO formula_headers (producto_nombre, lote_size_kg, activo) "
                 "VALUES ('PROD HUERF T1', 1, 1)")
    conn.execute("INSERT INTO formula_items (producto_nombre, material_id, material_nombre, porcentaje) "
                 "VALUES ('PROD HUERF T1', 'MPDEADCODE99', 'Ingrediente muerto', 100)")
    conn.execute("DELETE FROM maestro_mps WHERE codigo_mp='MPDEADCODE99'")
    conn.commit(); conn.close()
    xls = _excel_maestro([("X", "x", "MP-CR-OK", 1.0)], producto="OTRO PROD")
    r = c.post("/api/admin/cruce-maestro",
               data={"file": (io.BytesIO(xls), "m.xlsx")},
               headers=csrf_headers(), content_type="multipart/form-data")
    assert r.status_code == 200, r.data
    d = r.get_json()
    huerf = {h["producto"]: h for h in d.get("huerfanos", [])}
    assert "PROD HUERF T1" in huerf, d.get("huerfanos")
    assert "MPDEADCODE99" in huerf["PROD HUERF T1"]["codigos"], huerf["PROD HUERF T1"]
    assert huerf["PROD HUERF T1"]["en_excel"] is False


def test_archivar_producto_elimina_formula(app, db_clean):
    """Archivar borra formula_headers + formula_items del producto (descontinuado)."""
    c = _login(app)
    conn = sqlite3.connect(os.environ["DB_PATH"])
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp, nombre_inci, tipo_material, activo) "
                 "VALUES ('MP-ARCH-1', 'X', 'MP', 1)")
    conn.execute("INSERT INTO formula_headers (producto_nombre, lote_size_kg, activo) "
                 "VALUES ('PROD ARCHIVAR T1', 1, 1)")
    conn.execute("INSERT INTO formula_items (producto_nombre, material_id, material_nombre, porcentaje) "
                 "VALUES ('PROD ARCHIVAR T1', 'MP-ARCH-1', 'X', 100)")
    conn.commit(); conn.close()
    r = c.post("/api/admin/cruce-maestro/archivar-producto",
               json={"producto": "PROD ARCHIVAR T1"}, headers=csrf_headers())
    assert r.status_code == 200, r.data
    assert r.get_json()["items_eliminados"] == 1
    conn = sqlite3.connect(os.environ["DB_PATH"])
    n = conn.execute("SELECT COUNT(*) FROM formula_items WHERE producto_nombre='PROD ARCHIVAR T1'").fetchone()[0]
    nh = conn.execute("SELECT COUNT(*) FROM formula_headers WHERE producto_nombre='PROD ARCHIVAR T1'").fetchone()[0]
    conn.close()
    assert n == 0 and nh == 0, "la fórmula debe quedar eliminada"


def test_pares_clasifica_duplicado_vs_crossmap(app, db_clean):
    """El asistente clasifica: mismo INCI = DUPLICADO; INCI distinto = CROSS_MAP."""
    c = _login(app)
    conn = sqlite3.connect(os.environ["DB_PATH"])
    # par DUPLICADO: viejo y canónico con el MISMO INCI
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_inci,nombre_comercial,tipo_material,activo) VALUES ('MP-OLD-DUP','PANTHENOL','Pantenol viejo','MP',1)")
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_inci,nombre_comercial,tipo_material,activo) VALUES ('MP-CANON-DUP','PANTHENOL','Pantenol','MP',1)")
    # par CROSS_MAP: INCI distinto
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_inci,nombre_comercial,tipo_material,activo) VALUES ('MP-OLD-CROSS','ACETYL TETRAPEPTIDE-5','Pep viejo','MP',1)")
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_inci,nombre_comercial,tipo_material,activo) VALUES ('MP-CANON-CROSS','N-ACETYL GLUCOSAMINE','Glucosamina','MP',1)")
    # par DUPLICADO ES/EN: 'Tripeptido - 38' vs 'TRIPEPTIDE-38' = mismo (sinónimo idioma)
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_inci,nombre_comercial,tipo_material,activo) VALUES ('MP-OLD-ESEN','Palmitoyl Tripeptido - 38','Pep ES','MP',1)")
    conn.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_inci,nombre_comercial,tipo_material,activo) VALUES ('MP-CANON-ESEN','PALMITOYL TRIPEPTIDE-38','Pep EN','MP',1)")
    conn.execute("INSERT INTO formula_headers (producto_nombre,lote_size_kg,activo) VALUES ('PROD PARES T1',1,1)")
    # fórmula usa los códigos VIEJOS; el Excel dirá los canónicos (por nombre comercial)
    conn.execute("INSERT INTO formula_items (producto_nombre,material_id,material_nombre,porcentaje) VALUES ('PROD PARES T1','MP-OLD-DUP','Pantenol',5)")
    conn.execute("INSERT INTO formula_items (producto_nombre,material_id,material_nombre,porcentaje) VALUES ('PROD PARES T1','MP-OLD-CROSS','Glucosamina',5)")
    conn.execute("INSERT INTO formula_items (producto_nombre,material_id,material_nombre,porcentaje) VALUES ('PROD PARES T1','MP-OLD-ESEN','Palmitoyl tripeptide-38',5)")
    conn.commit(); conn.close()
    # Excel: el producto con los códigos CANÓNICOS, match por nombre comercial
    xls = _excel_maestro([
        ("PANTHENOL", "Pantenol", "MP-CANON-DUP", 5.0),
        ("N-ACETYL GLUCOSAMINE", "Glucosamina", "MP-CANON-CROSS", 5.0),
        ("PALMITOYL TRIPEPTIDE-38", "Palmitoyl tripeptide-38", "MP-CANON-ESEN", 5.0),
    ], producto="PROD PARES T1")
    r = c.post("/api/admin/cruce-maestro/pares",
               data={"file": (io.BytesIO(xls), "m.xlsx")},
               headers=csrf_headers(), content_type="multipart/form-data")
    assert r.status_code == 200, r.data
    pares = {p["old"]: p for p in r.get_json()["pares"]}
    assert pares.get("MP-OLD-DUP", {}).get("clase") == "DUPLICADO", pares
    assert pares.get("MP-OLD-CROSS", {}).get("clase") == "CROSS_MAP", pares
    # ES/EN: 'Tripeptido-38' == 'TRIPEPTIDE-38' → DUPLICADO (no cross-map)
    assert pares.get("MP-OLD-ESEN", {}).get("clase") == "DUPLICADO", pares


def test_cruce_maestro_requiere_admin(app, db_clean):
    c = _login(app, "jefferson")  # marketing, no admin
    xls = _excel_maestro([("X", "x", "MP-CR-Z", 1.0)])
    r = c.post("/api/admin/cruce-maestro",
               data={"file": (io.BytesIO(xls), "m.xlsx")},
               headers=csrf_headers(),
               content_type="multipart/form-data")
    assert r.status_code in (401, 403), r.data
