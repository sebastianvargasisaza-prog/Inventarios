"""Verificar fórmulas de la app contra el Excel maestro (2-jun-2026).

El Excel maestro es la fuente de verdad: tiene CÓD. BATCH por ingrediente.
Caso real: la app tiene 'N-acetil glucosamina' → MP00175 (tetrapéptido) pero el
Excel dice MP00262. El verificador lo detecta por NOMBRE y puede corregirlo."""
import os
import io
import sqlite3
import base64

import openpyxl

from .conftest import TEST_PASSWORD, csrf_headers


def _login(app):
    c = app.test_client()
    r = c.post("/login", data={"username": "sebastian", "password": TEST_PASSWORD},
               headers=csrf_headers(), follow_redirects=False)
    assert r.status_code == 302
    return c


def _excel_b64():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Contorno Test")
    ws.append(["#", "NOMBRE INCI", "NOMBRE COMERCIAL", "CÓD. BATCH", "% FÓRMULA", "g / 1 kg"])
    ws.append(["1", "N-ACETYL GLUCOSAMINE", "N-acetil glucosamina", "MP00262", "0.01", "10"])
    ws.append(["2", "GLYCERIN", "Glicerina", "MP00195", "0.03", "30"])
    bio = io.BytesIO()
    wb.save(bio)
    return base64.b64encode(bio.getvalue()).decode()


def _seed():
    db = sqlite3.connect(os.environ["DB_PATH"])
    try:
        for cod, nom in [("MP00262", "N-acetil glucosamina"), ("MP00195", "Glicerina"),
                         ("MP00175", "Acetyl tetrapeptide-5")]:
            db.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp, nombre_comercial, activo) VALUES (?,?,1)",
                       (cod, nom))
        db.execute("INSERT OR REPLACE INTO formula_headers (producto_nombre, unidad_base_g, lote_size_kg) "
                   "VALUES ('CONTORNO TEST', 1000, 1)")
        db.execute("DELETE FROM formula_items WHERE producto_nombre='CONTORNO TEST'")
        # glucosamina MAL mapeada a MP00175 (debería ser MP00262)
        db.execute("INSERT INTO formula_items (producto_nombre, material_id, material_nombre, porcentaje, cantidad_g_por_lote) "
                   "VALUES ('CONTORNO TEST','MP00175','N-acetil glucosamina',0.01,10)")
        db.execute("INSERT INTO formula_items (producto_nombre, material_id, material_nombre, porcentaje, cantidad_g_por_lote) "
                   "VALUES ('CONTORNO TEST','MP00195','Glicerina',0.03,30)")
        db.commit()
    finally:
        db.close()


def _clean():
    db = sqlite3.connect(os.environ["DB_PATH"])
    try:
        db.execute("DELETE FROM formula_items WHERE producto_nombre='CONTORNO TEST'")
        db.execute("DELETE FROM formula_headers WHERE producto_nombre='CONTORNO TEST'")
        db.commit()
    finally:
        db.close()


def test_verifica_detecta_codigo_mal(app, db_clean):
    _seed()
    try:
        c = _login(app)
        r = c.post("/api/admin/verificar-formulas-excel",
                   json={"contenido_b64": _excel_b64()}, headers=csrf_headers())
        assert r.status_code == 200, r.data
        d = r.get_json()
        rep = next((p for p in d["reporte"] if p["producto"] == "CONTORNO TEST"), None)
        assert rep, d
        mm = next((m for m in rep["mismatches"] if m["nombre"] == "N-acetil glucosamina"), None)
        assert mm, rep
        assert mm["codigo_app"] == "MP00175"
        assert mm["codigo_excel"] == "MP00262"
    finally:
        _clean()


def test_aplica_corrige_codigo(app, db_clean):
    _seed()
    try:
        c = _login(app)
        r = c.post("/api/admin/verificar-formulas-excel?aplicar=1",
                   json={"contenido_b64": _excel_b64()}, headers=csrf_headers())
        assert r.status_code == 200, r.data
        d = r.get_json()
        assert d["total_corregidos"] >= 1, d
        db = sqlite3.connect(os.environ["DB_PATH"])
        try:
            row = db.execute("SELECT material_id FROM formula_items WHERE producto_nombre='CONTORNO TEST' AND material_nombre='N-acetil glucosamina'").fetchone()
        finally:
            db.close()
        assert row[0] == "MP00262", row
    finally:
        _clean()


def _excel_b64_dos():
    import io as _io, base64 as _b64
    wb = openpyxl.Workbook(); wb.remove(wb.active); ws = wb.create_sheet("Dos Test")
    ws.append(["#", "NOMBRE INCI", "NOMBRE COMERCIAL", "CÓD. BATCH", "% FÓRMULA"])
    # glucosamina: excel MPGLUOK pero app la tiene en MPGLUSTOCK (que TIENE stock) -> no debe tocar en seguro
    ws.append(["1", "N-ACETYL GLUCOSAMINE", "N-acetil glucosamina", "MPGLUOK", "1"])
    # boron: excel MPBOROK (tiene stock), app en MPBORVACIO (0) -> SI debe corregir en seguro
    ws.append(["2", "BORON NITRIDE", "Boron nitride", "MPBOROK", "2"])
    bio = _io.BytesIO(); wb.save(bio); return _b64.b64encode(bio.getvalue()).decode()


def test_aplicar_seguro_solo_desbloquea(app, db_clean):
    db = sqlite3.connect(os.environ["DB_PATH"])
    try:
        for cod in ("MPGLUOK", "MPGLUSTOCK", "MPBOROK", "MPBORVACIO"):
            db.execute("INSERT OR REPLACE INTO maestro_mps (codigo_mp,nombre_comercial,activo) VALUES (?,?,1)", (cod, cod))
        # stock: MPGLUSTOCK tiene 500 (la app la usa, funciona) · MPGLUOK vacío
        db.execute("INSERT INTO movimientos (material_id,material_nombre,cantidad,tipo,fecha,lote,estado_lote) VALUES ('MPGLUSTOCK','glu',500,'Entrada','2026-06-01','LG','VIGENTE')")
        # boron: MPBOROK tiene 300 · MPBORVACIO 0 (la app la usa, bloqueada)
        db.execute("INSERT INTO movimientos (material_id,material_nombre,cantidad,tipo,fecha,lote,estado_lote) VALUES ('MPBOROK','bor',300,'Entrada','2026-06-01','LB','VIGENTE')")
        db.execute("INSERT OR REPLACE INTO formula_headers (producto_nombre,unidad_base_g,lote_size_kg) VALUES ('DOS TEST',1000,1)")
        db.execute("DELETE FROM formula_items WHERE producto_nombre='DOS TEST'")
        db.execute("INSERT INTO formula_items (producto_nombre,material_id,material_nombre,porcentaje,cantidad_g_por_lote) VALUES ('DOS TEST','MPGLUSTOCK','N-acetil glucosamina',1,10)")
        db.execute("INSERT INTO formula_items (producto_nombre,material_id,material_nombre,porcentaje,cantidad_g_por_lote) VALUES ('DOS TEST','MPBORVACIO','Boron nitride',2,20)")
        db.commit()
    finally:
        db.close()
    c = _login(app)
    r = c.post("/api/admin/verificar-formulas-excel?aplicar=1&solo_desbloquea=1",
               json={"contenido_b64": _excel_b64_dos()}, headers=csrf_headers())
    assert r.status_code == 200, r.data
    db = sqlite3.connect(os.environ["DB_PATH"])
    try:
        glu = db.execute("SELECT material_id FROM formula_items WHERE producto_nombre='DOS TEST' AND material_nombre='N-acetil glucosamina'").fetchone()[0]
        bor = db.execute("SELECT material_id FROM formula_items WHERE producto_nombre='DOS TEST' AND material_nombre='Boron nitride'").fetchone()[0]
    finally:
        db.close()
    assert glu == "MPGLUSTOCK", f"glucosamina NO debía tocarse (funciona) · quedó {glu}"
    assert bor == "MPBOROK", f"boron SÍ debía corregirse (estaba en 0) · quedó {bor}"
    db = sqlite3.connect(os.environ["DB_PATH"])
    try:
        db.execute("DELETE FROM formula_items WHERE producto_nombre='DOS TEST'")
        db.execute("DELETE FROM formula_headers WHERE producto_nombre='DOS TEST'")
        db.execute("DELETE FROM movimientos WHERE material_id IN ('MPGLUSTOCK','MPBOROK')")
        db.execute("DELETE FROM maestro_mps WHERE codigo_mp IN ('MPGLUOK','MPGLUSTOCK','MPBOROK','MPBORVACIO')")
        db.commit()
    finally:
        db.close()
