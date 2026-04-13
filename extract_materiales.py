#!/usr/bin/env python3
"""
Extrae lista maestra de materiales de todas las Fórmulas Maestras
Ejecutar: python3 extract_materiales.py
"""

import openpyxl
import os
from pathlib import Path

base_path = Path("Formulas Maestras")
materiales_dict = {}  # {nombre_upper: nombre_original}
contador = 0

print("🔍 Extrayendo materiales...\n")

# Procesar cada producto
for prod_folder in sorted(base_path.iterdir()):
    if not prod_folder.is_dir():
        continue

    # Buscar Excel
    for excel_file in prod_folder.glob("*.xlsx"):
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)

            if 'OP' not in wb.sheetnames:
                wb.close()
                continue

            ws = wb['OP']
            count_prod = 0

            # Extraer materiales desde fila 15
            for row in range(15, 150):
                cell_value = ws[f'B{row}'].value

                if cell_value is None:
                    break

                mat_name = str(cell_value).strip()
                if mat_name and mat_name != '':
                    mat_upper = mat_name.upper()

                    if mat_upper not in materiales_dict:
                        contador += 1
                        materiales_dict[mat_upper] = (mat_name, contador)
                        count_prod += 1

            if count_prod > 0:
                print(f"✓ {prod_folder.name:50} +{count_prod} nuevos")

            wb.close()

        except Exception as e:
            print(f"✗ {prod_folder.name:50} ({str(e)[:25]})")

print(f"\n{'='*70}")
print(f"✅ TOTAL MATERIALES ÚNICOS: {len(materiales_dict)}")
print(f"{'='*70}\n")

# Generar SQL
sql_lines = []
for mat_upper in sorted(materiales_dict.keys()):
    mat_original, seq = materiales_dict[mat_upper]
    codigo = f"MPMP{seq:05d}"
    # Escapar comillas simples
    mat_sql = mat_original.replace("'", "''")
    sql_lines.append(f"  ('{codigo}', '{mat_sql}', 'KG', TRUE)")

# Escribir SQL
sql_content = """-- LISTA MAESTRA MATERIALES - CÓDIGOS NORMALIZADOS
-- Total: """ + str(len(materiales_dict)) + """ materiales únicos

INSERT INTO materiales (codigo, nombre_inci, unidad, activo)
VALUES
""" + ",\n".join(sql_lines) + ";"

with open("FASE_6_LOAD_MATERIALES_LIMPIOS_FINAL.sql", "w", encoding="utf-8") as f:
    f.write(sql_content)

# Escribir lista de referencia
with open("LISTA_MAESTRA_MATERIALES.txt", "w", encoding="utf-8") as f:
    f.write("ID    | CÓDIGO     | NOMBRE\n")
    f.write("="*70 + "\n")
    for mat_upper in sorted(materiales_dict.keys()):
        mat_original, seq = materiales_dict[mat_upper]
        f.write(f"{seq:5d} | MPMP{seq:05d} | {mat_original}\n")

# Mostrar resultado
print("📋 MUESTRA (primeros 30 materiales):\n")
for mat_upper in sorted(materiales_dict.keys())[:30]:
    mat_original, seq = materiales_dict[mat_upper]
    print(f"  MPMP{seq:05d} | {mat_original}")

print(f"\n✅ Archivos generados:")
print(f"   • FASE_6_LOAD_MATERIALES_LIMPIOS_FINAL.sql")
print(f"   • LISTA_MAESTRA_MATERIALES.txt")
print(f"\nTotal: {len(materiales_dict)} materiales únicos")
