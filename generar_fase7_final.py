#!/usr/bin/env python3
"""
Genera FASE_7_CARGAR_FORMULAS_LIMPIAS.sql
Mapea cada ingrediente de formulas maestras a código MPMP
"""

from pathlib import Path
import openpyxl

print("=" * 80)
print("FASE 7: GENERANDO INSERT STATEMENTS DE FORMULAS CON CÓDIGOS MPMP")
print("=" * 80)

# Crear mapeo material -> MPMP code basado en LISTA_MAESTRA_CONSOLIDADA.txt
mapeo = {
    '1,2 HEXANEDIOL': 'MPMP00001',
    '3-O ACIDO ETIL ASCORBICO': 'MPMP00002',
    'ACEITE ARBOL DE TE': 'MPMP00003',
    'ACEITE DE ARGAN': 'MPMP00004',
    'ACEITE DE JOJOBA': 'MPMP00005',
    'ACEITE DE ROSA MOSQUETA': 'MPMP00006',
    'ACEITE JOJOBA': 'MPMP00007',
    'ACETYL HEXAPEPTIDE 8': 'MPMP00008',
    'ACETYL HEXAPEPTIDE-8': 'MPMP00009',
    'ACETYL HEXAPEPTIDO-8': 'MPMP00010',
    'ACETYL TETRAPEPTIDE 5': 'MPMP00011',
    'ACETYL TETRAPEPTIDE-3': 'MPMP00012',
    'ACETYL TETRAPEPTIDE-40': 'MPMP00013',
    'ACETYL TETRAPEPTIDE-40 (SALICYLIC ACID)': 'MPMP00014',
    'ACETYL TETRAPEPTIDE-5': 'MPMP00015',
    'ACETYL TETRAPETIDE-5': 'MPMP00016',
    'ACIDO ASCORBICO': 'MPMP00017',
    'ACIDO AZELAICO': 'MPMP00018',
    'ACIDO CAPRILOIL SALICILICO': 'MPMP00019',
    'ACIDO CITRICO': 'MPMP00020',
    'ACIDO GLUTAMICO': 'MPMP00021',
    'ACIDO HIALURONICO 1500 KD': 'MPMP00022',
    'ACIDO HIALURONICO 1500 KD (LYPHAR)': 'MPMP00023',
    'ACIDO HIALURONICO 1500KD': 'MPMP00024',
    'ACIDO HIALURONICO 300 KD': 'MPMP00025',
    'ACIDO HIALURONICO 300 KD (LYPHAR)': 'MPMP00026',
    'ACIDO HIALURONICO 300KD': 'MPMP00027',
    'ACIDO HIALURONICO 50 KD': 'MPMP00028',
    'ACIDO HIALURONICO 50 KD (LYPHAR)': 'MPMP00029',
    'ACIDO HIALURONICO 50KD': 'MPMP00030',
    'ACIDO HIALURÓNICO 1500 KD': 'MPMP00031',
    'ACIDO HIALURÓNICO 300 KD': 'MPMP00032',
    'ACIDO HIALURÓNICO 50 KD': 'MPMP00033',
    'ACIDO HILAURONICO 50 KD': 'MPMP00034',
    'ACIDO KOJICO': 'MPMP00035',
    'ACIDO LACTICO': 'MPMP00036',
    'ACIDO SALICILICO': 'MPMP00037',
    'ACIDO TRANEXAMICO': 'MPMP00038',
    'ADENOSINA': 'MPMP00039',
    'ADIPATO DE BUTIL': 'MPMP00040',
    'AGUA DESIONIZADA': 'MPMP00041',
    'ALANTOINA': 'MPMP00042',
    'ALANTOÍNA': 'MPMP00043',
    'ALCOHOL CETILICO': 'MPMP00044',
    'ALCOHOL CETITLICO': 'MPMP00045',
    'ALFA  ARBUTINA': 'MPMP00046',
    'ALFA ARBUTINA': 'MPMP00047',
    'ALFA ARBUTINA  (LYPHAR)': 'MPMP00048',
    'ALFA OLEFIN SULFONATO DE SODIO': 'MPMP00049',
    'ALOE VERA': 'MPMP00050',
    'ANEMARRHENA ASPHODELOIDES ROOT EXTRACT': 'MPMP00051',
    'AOS 40': 'MPMP00052',
    'ARGANIA SPINOSA KERNEL OIL': 'MPMP00053',
    'ASCORBATO DE TETRAHEXILDECILO': 'MPMP00054',
    'ASCORBIC GLUCOSIDE': 'MPMP00055',
    'ASCORBIL GLUCOSIDE': 'MPMP00056',
    'ASCORBYL GLUCOSIDE': 'MPMP00057',
    'ASIATICOSIDO': 'MPMP00058',
    'ASTAXANTINA': 'MPMP00059',
    'AZELOLIL DIGLICINATO DE POTASIO': 'MPMP00060',
    'BACKUCHIOL': 'MPMP00061',
    'BAKUCHIOL': 'MPMP00062',
    'BENZOATO DE SODIO': 'MPMP00063',
    'BETA-CICLODEXTRINA': 'MPMP00064',
    'BETAGLUCAM': 'MPMP00065',
    'BETAGLUCAN': 'MPMP00066',
    'BETAINA': 'MPMP00067',
    'BETAINE SALICYLATE': 'MPMP00068',
    'BICARBONATO DE SODIO': 'MPMP00069',
    'BICARBONATO SODIO': 'MPMP00070',
    'BIOSURE FE': 'MPMP00071',
    'BIOTYNOIL TRIPEPTIDO-1': 'MPMP00072',
    'BISABOLOL': 'MPMP00073',
    'CAFEINA': 'MPMP00074',
    'CAPRYLOYL SALICYLIC ACID': 'MPMP00075',
    'CARBOPOL': 'MPMP00076',
    'CENTELLA': 'MPMP00077',
    'CENTELLA ASIATICA POLVO': 'MPMP00078',
    'CHITOSAN': 'MPMP00079',
    'COCOS NUCIFERA OIL': 'MPMP00080',
    'COOPER TRIPEPTIDE-1': 'MPMP00081',
    'COPPER TRIPEPTIDE 1': 'MPMP00082',
    'COPPER TRIPEPTIDO': 'MPMP00083',
    'DIPEPTIDE DIAMINOBUTYROYL BENZYLAMIDE DIACETTE': 'MPMP00084',
    'DIPÉPTIDO DIAMINOBUTIROIL BENZALAMIDA DIACETATO': 'MPMP00085',
    'DISODIUM EDTA': 'MPMP00086',
    'ECTOINA': 'MPMP00087',
    'EDTA DISODICO': 'MPMP00088',
    'EPI-ON (AZELAMIDOPROPIL DIMETIL AMINA)': 'MPMP00089',
    'ERGOTIONEINA': 'MPMP00090',
    'ESCUALENO': 'MPMP00091',
    'EZ-4U': 'MPMP00092',
    'FENOXIETANOL': 'MPMP00093',
    'FITATO DE SODIO': 'MPMP00094',
    'FOSFATO DE ASCORBILO SODICO': 'MPMP00095',
    'FRAGANCIA FRESA CREMOSO': 'MPMP00096',
    'FRAGANCIA PISTACHO': 'MPMP00097',
    'FRAGANCIA YOGURT CREMOSO': 'MPMP00098',
    'GIGA WHITE': 'MPMP00099',
    'GLICERINA': 'MPMP00100',
    'GLICINA': 'MPMP00101',
    'GLICINAMIDA': 'MPMP00102',
    'GLUCONOLACTONA': 'MPMP00103',
    'GLUCOSAMINA (NAG)': 'MPMP00104',
    'GLUCOSAMINA NAG': 'MPMP00105',
    'GLUTATION': 'MPMP00106',
    'GLUTATIÓN': 'MPMP00107',
    'GOMA XANTAN': 'MPMP00108',
    'GRANSIL VX 419': 'MPMP00109',
    'GRANSIL VX419': 'MPMP00110',
    'HDI/TRIMETHYLOL HEXYLLACTONE CROSSPOLYMER (AND) SILICA': 'MPMP00111',
    'HIDROXIDO DE SODIO (SODA CAUSTICA)': 'MPMP00112',
    'HIDROXIDO DE SODIO (SOLUCION 50%)': 'MPMP00113',
    'HIDROXIDO SODIO': 'MPMP00114',
    'HYDROLYZED RICE EXTRACT': 'MPMP00115',
    'HYDROXYPINACOLONE RETINOATE': 'MPMP00116',
    'L-CARNITINA': 'MPMP00117',
    'L-TEANINA': 'MPMP00118',
    'LACTOBIONIC ACID': 'MPMP00119',
    'LAURIL GLUCOSIDO': 'MPMP00120',
    'LAURYL GLUCOSIDE': 'MPMP00121',
    'LAVANDULA ANGUSTIFOLIA OIL': 'MPMP00122',
    'MELATONIN': 'MPMP00123',
    'MELATONINA': 'MPMP00124',
    'METILSULFONILMETANO': 'MPMP00125',
    'MYRISTOIL NONAPEPTIDE-3': 'MPMP00126',
    'MYRISTOYL HEXAPEPTIDO-16': 'MPMP00127',
    'MYRISTOYL PENTAPEPTIDE-17': 'MPMP00128',
    'N-ACETIL GLUCOSAMINA': 'MPMP00129',
    'N-ACETILGLUCOSAMINA': 'MPMP00130',
    'N-ACETL GLUCOSAMINA': 'MPMP00131',
    'N-ACETYL-L-CYSTEINE': 'MPMP00132',
    'NAD': 'MPMP00133',
    'NIACINAMIDA': 'MPMP00134',
    'NMN': 'MPMP00135',
    'OLIGOPEPTIDO 68': 'MPMP00136',
    'OLIGOPEPTIDO-68': 'MPMP00137',
    'OLIVE OIL PEG-7-ESTERS': 'MPMP00138',
    'PALMITATO DE ISOPROPILO': 'MPMP00139',
    'PALMITOYL PENTAPEPTIDE 4': 'MPMP00140',
    'PALMITOYL PENTAPEPTIDE-4': 'MPMP00141',
    'PALMITOYL TETRAPEPTIDE 7': 'MPMP00142',
    'PALMITOYL TETRAPEPTIDE-7': 'MPMP00143',
    'PALMITOYL TETRAPEPTIDO-7': 'MPMP00144',
    'PALMITOYL TRIPEPTIDE 1': 'MPMP00145',
    'PALMITOYL TRIPEPTIDE 5': 'MPMP00146',
    'PALMITOYL TRIPEPTIDE-1': 'MPMP00147',
    'PALMITOYL TRIPEPTIDE -5': 'MPMP00148',
    'PALMITOYL TRIPEPTIDO -5': 'MPMP00148',
    'PALMITOYL TRIPEPTIDO-1': 'MPMP00149',
    'PALMITOYL TRIPEPTIDO-38': 'MPMP00150',
    'PALMITOYL TRIPETIDE-5': 'MPMP00151',
    'PALMIYOL TETRAPEPTIDE-7': 'MPMP00152',
    'PALMIYOL TRIPEPTIDE-1': 'MPMP00153',
    'PALMIYOL TRIPEPTIDE-5': 'MPMP00154',
    'PANTENOL - LIQUIDO': 'MPMP00155',
    'PANTENOL LIQUIDO': 'MPMP00156',
    'PANTENOL POLVO': 'MPMP00157',
    'PDRN (SODIUM DNA)': 'MPMP00158',
    'PEPTIDOS DE COLAGENO': 'MPMP00159',
    'PEPTIDOS DE COLAGENO POLVO': 'MPMP00160',
    'POLIETILENGLICOL 400': 'MPMP00161',
    'POTASSIUM SORBATE': 'MPMP00162',
    'PROBETAINA': 'MPMP00163',
    'PROLINE': 'MPMP00164',
    'PROPILENGLICOL': 'MPMP00165',
    'QUIMCREAM': 'MPMP00166',
    'QUINCREAM': 'MPMP00167',
    'REGALIZ': 'MPMP00168',
    'RESVERATROL': 'MPMP00169',
    'RETINALDEHIDO': 'MPMP00170',
    'RETINALDEHÍDO': 'MPMP00171',
    'RETINOL 99%': 'MPMP00172',
    'RETINYL RETINOATE': 'MPMP00173',
    'ROSMARINUS OFFICINALIS LEAF OIL': 'MPMP00174',
    'SILICONA BM 600': 'MPMP00175',
    'SILICONA LIQUIDA': 'MPMP00176',
    'SILIMARINA': 'MPMP00177',
    'SODIUM COCOYL GLICYNATE': 'MPMP00178',
    'SODIUM LAUROYL SARCOSINATE': 'MPMP00179',
    'SODIUM PHYTATE': 'MPMP00180',
    'SORBATO DE POTASIO': 'MPMP00181',
    'TERPENOS SOLUBLE': 'MPMP00182',
    'TERPENOS SOLUBLES': 'MPMP00183',
    'TERPENOS SOLUBLES 80%': 'MPMP00184',
    'TERPENOS SOLUBLES 98%': 'MPMP00185',
    'TINOGARD TT': 'MPMP00186',
    'TRANEXAMIC ACID': 'MPMP00187',
    'TRIETANOLAMINA 85%': 'MPMP00188',
    'TRIFOLIUM PRATENSE FLOWER EXTRACT': 'MPMP00189',
    'TRIGLICERIDO CAPRICO': 'MPMP00190',
    'TWEEN 20': 'MPMP00191',
    'TWEEN 80': 'MPMP00192',
    'UNDECILENOIL FENILALANINA': 'MPMP00193',
    'UREA': 'MPMP00194',
    'VITAMINA E  POLVO': 'MPMP00195',
    'VITAMINA E - ACEITE': 'MPMP00196',
    'VITAMINA E ACEITE': 'MPMP00197',
    'VITAMINA E LIQUIDA': 'MPMP00198',
    'VITAMINA E POLVO': 'MPMP00199',
    'ZINC PCA': 'MPMP00200',
    'ÁCIDO AZELAICO': 'MPMP00201',
    'ÁCIDO FERÚLICO': 'MPMP00202',
    'ÁCIDO HIALURONICO 50 KD': 'MPMP00203',
    'ÁCIDO HIALURÓNICO 1500 KD': 'MPMP00204',
    'ÁCIDO HIALURÓNICO 300 KD': 'MPMP00205',
    'ÁCIDO HIALURÓNICO 50 KD': 'MPMP00206',
    'ÁCIDO KOJICO': 'MPMP00207',
    'ÁCIDO LACTICO': 'MPMP00208',
    'ÁCIDO LÁCTICO': 'MPMP00209',
    'ÁCIDO SALICILICO': 'MPMP00210',
    'ÁCIDO TRANEXAMICO': 'MPMP00211',
    'ÚREA': 'MPMP00212',
}

print(f"\n✓ Mapeo cargado: {len(mapeo)} materiales")

# Leer fórmulas maestras
print("\nLeyendo formulas maestras...")
formulas_data = {}
base = Path("Formulas Maestras")
producto_count = 0

for prod_folder in sorted(base.iterdir()):
    if not prod_folder.is_dir():
        continue

    prod_name = prod_folder.name

    for excel_file in sorted(prod_folder.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            if 'OP' in wb.sheetnames:
                ws = wb['OP']

                ingredientes = []
                for row in range(15, 100):
                    nombre_cell = ws[f'B{row}'].value
                    cantidad_cell = ws[f'C{row}'].value

                    if nombre_cell:
                        nombre = str(nombre_cell).strip().upper()
                        cantidad = float(cantidad_cell) if cantidad_cell else 0

                        # Buscar en mapeo con normalización
                        mpmp_code = mapeo.get(nombre)

                        if mpmp_code and cantidad > 0:
                            ingredientes.append((mpmp_code, cantidad))
                    elif nombre_cell is None:
                        break

                if ingredientes:
                    if prod_name not in formulas_data:
                        formulas_data[prod_name] = []
                    formulas_data[prod_name].extend(ingredientes)
                    producto_count += 1

            wb.close()
        except Exception as e:
            print(f"  ⚠️  Error en {excel_file.name}: {e}")

print(f"✓ {len(formulas_data)} productos con fórmulas")

# Generar SQL
print("\nGenerando SQL INSERT...")
sql_lines = []

for producto in sorted(formulas_data.keys()):
    for mpmp_code, cantidad in sorted(formulas_data[producto]):
        sql_lines.append(f"  ('{mpmp_code}', '{producto}', {cantidad})")

# Escribir archivo
sql_content = f"""-- ============================================================================
-- FASE 7: CARGAR FORMULAS LIMPIAS CON CÓDIGOS MPMP NORMALIZADOS
-- ============================================================================
-- Tabla: formulas_productos
-- Columnas: codigo_material (FK materiales.codigo), nombre_producto, cantidad
-- Total relaciones: {len(sql_lines)}
-- ============================================================================

INSERT INTO formulas_productos (codigo_material, nombre_producto, cantidad)
VALUES
""" + ",\n".join(sql_lines) + ";"

with open("FASE_7_CARGAR_FORMULAS_LIMPIAS.sql", "w", encoding="utf-8") as f:
    f.write(sql_content)

print(f"\n✅ ARCHIVO GENERADO:")
print(f"   📄 FASE_7_CARGAR_FORMULAS_LIMPIAS.sql")
print(f"\n📊 ESTADÍSTICAS:")
print(f"   • Productos: {len(formulas_data)}")
print(f"   • Relaciones ingrediente-producto: {len(sql_lines)}")
print(f"   • Materiales mapeados: {len(mapeo)}")
print("=" * 80)
